"""Golden checks for the audit-round hardening (compare_core.py + compare_env.py).

Locks the fixes found in the ruthless multi-agent audit (all latent on real data
but real invariant/robustness gaps):
  P1  the data-sheet HELPER-KEY cell and the Routes ROUTE-ID cell are guarded —
      a "="-leading key/route no longer becomes a live formula (which broke the
      Comparison MATCH lookups AND split the two flavors). [completes injection fix]
  P2  hidden, versioned state masks own discrepancy truth. Diffs, Summary,
      conditional formatting, and Spot Check consume E/D/N/U state or an exact
      state/display twin; the visible " ≠ " separator is presentation only.
  P3  Ramp Summary route keys are zero-pad-normalized (PDF title "1" == filename
      "001"); internal unnamed header columns get an identifiable placeholder.
  P4  normalize_value canonicalizes datetime.time too.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_audit.py
"""
import os
import sys
import tempfile
from datetime import time
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env as env
from compare_core import CompareSchema, normalize_value, run_compare
from openpyxl import Workbook, load_workbook


def _col(ws, name):
    for c in next(ws.iter_rows(max_row=1)):
        if c.value == name:
            return c.column
    raise AssertionError(f"{name!r} not in {ws.title}")


def test_p1_helper_key_and_route_id_guarded():
    sc = CompareSchema(report_name="K", header=["Loc", "V"], side_a="A", side_b="B",
                       id_noun="row", id_noun_plural="rows")
    out = os.path.join(tempfile.gettempdir(), "_audit_hk.xlsx")
    # per-route: a "="-leading KEY. The data helper must be an opaque literal,
    # never source-derived formula text (else MATCH lookups break + flavors split).
    run_compare(sc, [["=K1", "x"], ["b", "y"]], [["=K1", "x"], ["b", "y"]],
                False, out, mode="formulas")
    wb = load_workbook(out, data_only=False)
    ws = wb["A"]
    hk_col = _col(ws, "Key (helper)")
    hk = ws.cell(row=2, column=hk_col)
    assert (hk.data_type == "s"
            and str(hk.value).startswith("__CMP_E2_KEY_V1_")), \
        (hk.data_type, hk.value)
    raw_key = ws.cell(row=2, column=_col(ws, "Loc"))
    assert raw_key.data_type == "s" and raw_key.value == "=K1", \
        (raw_key.data_type, raw_key.value)
    wb.close(); os.remove(out)

    # consolidated: a "="-leading ROUTE id on the Routes sheet must be TEXT.
    scr = CompareSchema(report_name="K", header=["Loc", "V"], side_a="A", side_b="B",
                        id_noun="row", id_noun_plural="rows")
    run_compare(scr, [["=R1", "1", "x"]], [["=R1", "1", "x"]], True, out,
                mode="formulas")
    wb = load_workbook(out, data_only=False)
    rid = wb["Routes"].cell(row=2, column=1)
    assert rid.data_type == "s" and rid.value == "=R1", (rid.data_type, rid.value)
    wb.close(); os.remove(out)


def test_p2_state_mask_is_truth():
    sc = CompareSchema(report_name="K", header=["Loc", "V"], side_a="A", side_b="B",
                       id_noun="row", id_noun_plural="rows")
    out = os.path.join(tempfile.gettempdir(), "_audit_cf.xlsx")
    run_compare(sc, [["1", "a"]], [["1", "b"]], False, out, mode="formulas")
    wb = load_workbook(out, data_only=False)
    try:
        comparison = wb["Comparison"]
        state_cells = [cell for cell in comparison[1]
                       if str(cell.value).startswith("__CMP_E1_STATE_V1_")]
        assert len(state_cells) == 1, [cell.value for cell in state_cells]
        state_col = state_cells[0].column_letter
        assert comparison.column_dimensions[state_col].hidden, state_col
        state_formula = str(comparison[f"{state_col}2"].value)
        assert '"E"' in state_formula and '"D"' in state_formula, state_formula

        diffs = str(comparison.cell(2, _col(comparison, "Diffs")).value)
        summary_formulas = [str(cell.value) for row in wb["Summary"].iter_rows()
                            for cell in row if cell.data_type == "f"]
        per_field = [formula for formula in summary_formulas
                     if "MID(Comparison!" in formula and '="D"' in formula]
        cf_formulas = [str(formula)
                       for region in comparison.conditional_formatting
                       for rule in comparison.conditional_formatting[region]
                       for formula in (rule.formula or ())]
        spot = wb["Spot Check"]
        agreement = str(spot["G16"].value)

        assert "SUBSTITUTE(" in diffs and '"D"' in diffs, diffs
        assert per_field, summary_formulas
        assert any("MID($" in formula and '="D"' in formula
                   for formula in cf_formulas), cf_formulas
        assert ("EXACT($K16,$M16)" in agreement
                and "EXACT($L16,$F16)" in agreement), agreement
        assert all(spot.column_dimensions[col].hidden for col in ("K", "L", "M"))

        # The separator may appear in display formulas and prose, but no truth
        # consumer may scan for it. This rejects the old forged-state path.
        truth_formulas = [diffs, agreement, *per_field, *cf_formulas]
        marker_scans = [formula for formula in truth_formulas
                        if "≠" in formula
                        and any(token in formula.upper()
                                for token in ("SEARCH(", "FIND(", "COUNTIF("))]
        assert not marker_scans, marker_scans
    finally:
        wb.close(); os.remove(out)


def test_p5_spot_row_matching_independent():
    """CMP-AUD-218: Spot Check derives BOTH data-sheet rows by MATCHing the
    hidden Comparison key-token column into each side's literal "Key (helper)"
    column — never by reading Comparison's stored row links — and the
    Row-integrity line (C14) EXACT-compares Comparison's claimed links/status
    against that independent derivation. Pinned on both workbook twins."""
    sc = CompareSchema(report_name="K", header=["Loc", "V"], side_a="A", side_b="B",
                       id_noun="row", id_noun_plural="rows")
    out = os.path.join(tempfile.gettempdir(), "_audit_spot218.xlsx")
    values_out = os.path.join(tempfile.gettempdir(),
                              "_audit_spot218 (values).xlsx")
    run_compare(sc, [["1", "a"], ["2", "b"]], [["1", "x"], ["2", "b"]], False,
                out, mode="both")
    for path in (out, values_out):
        wb = load_workbook(path, data_only=False)
        try:
            comparison = wb["Comparison"]
            token_cells = [cell for cell in comparison[1]
                           if cell.value == "__CMP_E2_KEY_V1_TOKEN"]
            assert len(token_cells) == 1, [c.value for c in comparison[1]]
            token_col, token_letter = (token_cells[0].column,
                                       token_cells[0].column_letter)
            assert comparison.column_dimensions[token_letter].hidden, path
            tokens = [comparison.cell(row, token_col) for row in (2, 3)]
            assert all(cell.data_type == "s"
                       and str(cell.value).startswith("__CMP_E2_KEY_V1_")
                       for cell in tokens), \
                [(cell.data_type, cell.value) for cell in tokens]
            assert len({cell.value for cell in tokens}) == 2, \
                [cell.value for cell in tokens]

            spot = wb["Spot Check"]
            token_pull = str(spot["M12"].value)
            row_a, row_b = str(spot["K12"].value), str(spot["L12"].value)
            c12, f12 = str(spot["C12"].value), str(spot["F12"].value)
            callout = str(spot["B13"].value)
            integrity = str(spot["C14"].value)
            assert (f"INDEX(Comparison!${token_letter}:${token_letter}"
                    in token_pull), token_pull
            assert "MATCH($M$12," in row_a and "MATCH($M$12," in row_b, \
                (row_a, row_b)
            # The audited rows and the one-sided callout never consume
            # Comparison's stored links or status.
            for formula in (row_a, row_b, c12, f12, callout):
                assert "INDEX(Comparison!" not in formula, formula
            assert "$K$12" in c12 and "$L$12" in f12, (c12, f12)
            assert "$K$12" in callout and "$L$12" in callout \
                and "$C$11" not in callout, callout
            # The Row-integrity line is the ONLY consumer of the claimed
            # links: it EXACT-compares them against the independent match.
            assert ("EXACT(" in integrity and "$K$12" in integrity
                    and "$L$12" in integrity
                    and "INDEX(Comparison!" in integrity), integrity
        finally:
            wb.close()
    os.remove(out); os.remove(values_out)


def test_p3_route_key_normalize():
    assert env._norm_route_key("1") == "001"
    assert env._norm_route_key("001") == "001"
    assert env._norm_route_key("101U") == "101U"
    assert env._norm_route_key("5s") == "005S"
    assert env._norm_route_key("1") == env._norm_route_key("001")   # the bug case


def test_p3_none_header_placeholder():
    d = Path(tempfile.mkdtemp()) / "highway_sequence"
    d.mkdir(parents=True)
    wb = Workbook(); ws = wb.active; ws.title = "Highway Locations"
    ws.append(["County", "City", None, "PM", None, "Description"])  # internal Nones
    ws.append(["ORA", None, "R", "000.1", None, "JCT"])
    wb.save(d / "highway_sequence_route_001.xlsx")
    from events import Events
    rows, header, skipped = env._load_xlsx_side(
        d.parent, "X", "highway_sequence", "Highway Locations", "Highway Sequence",
        Events())
    assert None not in header, ("internal None must be relabeled", header)
    assert header[2] == "(col C)" and header[4] == "(col E)", header
    assert header[0] == "County" and header[3] == "PM", header


def test_p4_time_normalize():
    assert normalize_value(time(8, 30)) == "08:30:00"
    assert normalize_value(time(8, 30, 5, 123456)) == "08:30:05.123456"


def test_p3_side_label_cap():
    # Same run-folder name in different parents -> date + (A)/(B) -> >23 chars
    # -> must be capped so "Only in <label>" fits Excel's 31-char sheet limit.
    base = Path(tempfile.mkdtemp())
    a = base / "x1" / "2026-06-15 ssor-prod"; a.mkdir(parents=True)
    b = base / "x2" / "2026-06-15 ssor-prod"; b.mkdir(parents=True)
    la, lb = env._side_labels(a, b)
    assert la != lb, (la, lb)
    assert len("Only in " + la) <= 31 and len("Only in " + lb) <= 31, (la, lb)


def main():
    test_p1_helper_key_and_route_id_guarded()
    test_p2_state_mask_is_truth()
    test_p5_spot_row_matching_independent()
    test_p3_route_key_normalize()
    test_p3_none_header_placeholder()
    test_p4_time_normalize()
    test_p3_side_label_cap()
    print("OK  audit hardening: helper-key + route-id guarded; hidden state masks "
          "own diff truth with no marker scans; Spot Check row matching rides the "
          "hidden key token (never Comparison's links) with a Row-integrity line; "
          "route keys normalized; unnamed columns labeled; time canonicalized; "
          "side labels capped to the 31-char limit.")


if __name__ == "__main__":
    main()
