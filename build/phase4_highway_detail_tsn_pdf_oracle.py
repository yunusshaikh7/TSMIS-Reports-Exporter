"""Independent Highway Detail TSN XLSX <-> district-PDF parity oracle.

This audit intentionally does not import any application comparator, TSN loader,
evidence adapter, parser, report-view writer, or their constants.  The raw 56-
column workbook and the twelve vendor district prints are its only business-data
inputs.  Production modules may be inspected while developing this oracle, but
cannot make a failing source fact pass here.

The oracle proves three different things separately:

* source binding: exact hashes, workbook schema, PDF page/member metadata;
* conservation: every parsed two-line PDF record maps to one XLSX physical
  identity (route suffix + county + complete PP/PM + roadbed + equation claims),
  while exact duplicate multiplicity is retained;
* representation parity: every printable XLSX field is reconciled against its
  PDF line/section.  Differences are either documented rendering equivalence,
  exact seven-day source-snapshot delta, or unresolved residue.  Parser loss,
  ambiguous non-identical duplicates, and unclassified residue fail closed.

The JSON result contains the explicit XLSX-field -> PDF-section mapping used to
audit the layered Highway Detail Report View.
"""

from __future__ import annotations

import argparse
import collections
import datetime as dt
import difflib
import hashlib
import itertools
import json
import re
import sys
import time
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - exercised by the caller's runtime
    raise SystemExit(f"required audit dependency is missing: {exc}")


EXPECTED_XLSX_HEADER = (
    "THY_ID", "DIST", "CNTY", "RTE", "RTE_SFX", "DIST_CNTY_ROUTE", "PP",
    "POSTMILE", "E_IND", "LENGTH", "REC_DATE", "HG", "AC", "ACC_SIG",
    "ACC_EFF_DATE", "CITY", "POP_CODE", "BEG_DATE", "ADT_AMT", "PROFILE",
    "BREAK_DESC", "LK_BACK_ADT", "CHNGMILE", "DVM", "DESCRIPTION", "NON_ADD",
    "LT_SIG", "L_EFF_DATE", "L_ST", "L_NO_LANES", "L_SF", "L_OT_TOT",
    "L_OT_TR", "L_TR_WID", "L_IN_TOT", "L_IN_TR", "MED_SIG", "M_EFF_DATE",
    "M_TYPE_CODE", "M_CL", "M_BA", "M_WID", "M_VA", "RT_SIG", "R_EFF_DATE",
    "R_ST", "R_NO_LANES", "R_SF", "R_IN_TOT", "R_IN_TR", "R_TR_WID",
    "R_OT_TOT", "R_OT_TR", "SEG_ORDER_ID", "REFERENCE_DATE", "EXTRACT_DATE",
)

EXPECTED_XLSX_SHA256 = "bac3c882002b26433e39fad00c3dcdf9ad95b8dfc9ba9597386c656a71071dd1"
EXPECTED_XLSX_ROWS = 60_083
EXPECTED_PDF_MEMBERS = 12
EXPECTED_PDF_PAGES = 4_123
EXPECTED_XLSX_REFERENCE_DATE = "2025-09-08"
EXPECTED_XLSX_EXTRACT_DATE = "2025-09-15"
EXPECTED_PDF_REPORT_DATE = "2025-09-15"
EXPECTED_PDF_REFERENCE_DATE = "2025-09-15"

EXPECTED_PDF_SHA256 = {
    "D01 Highway Detail_TSN.pdf": "815ad64218ebcd262ceb75b8efa65c6173023f543ecef28beb33b096e7fb5ce4",
    "D02 Highway Detail_TSN.pdf": "7076f29147bc119cdf62398c99104ee7b7bfa05ee44663144a09303e805fd8ae",
    "D03 Highway Detail_TSN.pdf": "cdbfc4abda1e67cf16b46f43a3ae750e54aace7a0fa738b1cae66dedf30d3885",
    "D04 Highway Detail_TSN.pdf": "3fcb00d918b3cc2ce78043c50e01085e342c08cd9ba754a5ce9a7d074ae2ff5e",
    "D05 Highway Detail_TSN.pdf": "099460bc87ad6fb6c6fac00753be099e433cd54f7c77d636c7c8b97aebb42a04",
    "D06 Highway Detail_TSN.pdf": "7b4ec9888164c8ab2be424f13b6a77235dfd11a59e107183b1edbe2bc33d50d5",
    "D07 Highway Detail_TSN.pdf": "19538cc8569e7af578cc592d836a7f2cebfab99a2fc1b23498684fe92249c8ec",
    "D08 Highway Detail_TSN.pdf": "3507b0647a1a8851bb67b4a668a96c66697af2574da9eb68a7b66a8ea6aa8d46",
    "D09 Highway Detail_TSN.pdf": "d631142229b619bbadfe685aa6e944646757dc487e2e2051ed3891afacf25123",
    "D10 Highway Detail_TSN.pdf": "af0bc372674bbac9d6821a0f1faff246697523e02b67802f9a96bdfc637f05f6",
    "D11 Highway Detail_TSN.pdf": "a987c4a47e09a18462a16874780c3a06f0c3aa59ca71068e611b8e76b0979f25",
    "D12 Highway Detail_TSN.pdf": "381d895c242bf5606360477e4290383da7ac807b9bdd09b3818b33649b1d40a1",
}

# Filled only after a discovery run has enumerated the complete, field-specific
# delta set for the exact XLSX/PDF hashes above.  A normal full run refuses any
# addition, removal, value change, identity change, or field-count change.
EXPECTED_DELTA_MANIFEST_SHA256 = "d101bc1263188dcb436a9218bad6774ab047368e819c205d1e53b9b812b56d8a"
EXPECTED_DELTA_ITEM_COUNT = 443
EXPECTED_DELTA_FIELD_COUNTS = {
    "DESCRIPTION": 3,
    "LENGTH": 2,
    "LK_BACK_ADT": 436,
}


FIELD_TO_PDF = collections.OrderedDict([
    ("DIST", "DCR header: DIST-CNTY-ROUTE / district"),
    ("CNTY", "DCR header: DIST-CNTY-ROUTE / county"),
    ("RTE", "DCR header: DIST-CNTY-ROUTE / route base"),
    ("RTE_SFX", "DCR header: DIST-CNTY-ROUTE / route suffix"),
    ("DIST_CNTY_ROUTE", "DCR header: complete composite"),
    ("PP", "record line 1 / POST MILE prefix"),
    ("POSTMILE", "record line 1 / POST MILE numeric component"),
    ("E_IND", "record line 1 / P S equation marker"),
    ("LENGTH", "record line 1 / LENGTH"),
    ("REC_DATE", "record line 1 / DATE OF RECORD"),
    ("HG", "record line 1 / H G"),
    ("AC", "record line 1 / A C"),
    ("ACC_SIG", "record line 1 / ACC-CONT change flag"),
    ("ACC_EFF_DATE", "record line 1 / ACC-CONT EFF DATE"),
    ("CITY", "record line 1 / CITY CODE"),
    ("POP_CODE", "record line 1 / R U"),
    ("BEG_DATE", "record line 1 / ADT INFORMATION EFF-DATE (printed as count-year YY-01-01)"),
    ("ADT_AMT", "record line 1 / ADT INFORMATION LK-AHD"),
    ("PROFILE", "record line 1 / ADT INFORMATION P"),
    ("LK_BACK_ADT", "record line 1 / ADT INFORMATION LK-BACK"),
    ("CHNGMILE", "record line 1 / ADT INFORMATION CHANGE/MILE"),
    ("DVM", "record line 1 / ADT INFORMATION D V M"),
    ("DESCRIPTION", "record line 2 / DESCRIPTION (first 23 source characters)"),
    ("NON_ADD", "record line 2 / N A"),
    ("LT_SIG", "record line 2 / LEFT ROADBED change flag"),
    ("L_EFF_DATE", "record line 2 / LEFT ROADBED EFF-DATE"),
    ("L_ST", "record line 2 / LEFT ROADBED S T"),
    ("L_NO_LANES", "record line 2 / LEFT ROADBED NO LN"),
    ("L_SF", "record line 2 / LEFT ROADBED S F"),
    ("L_OT_TOT", "record line 2 / LEFT ROADBED OT-SH TO (3-character cell)"),
    ("L_OT_TR", "record line 2 / LEFT ROADBED OT-SH TR (3-character cell)"),
    ("L_TR_WID", "record line 2 / LEFT ROADBED T-W WID"),
    ("L_IN_TOT", "record line 2 / LEFT ROADBED IN-SH TO (3-character cell)"),
    ("L_IN_TR", "record line 2 / LEFT ROADBED IN-SH TR (3-character cell)"),
    ("MED_SIG", "record line 2 / MEDIAN change flag"),
    ("M_EFF_DATE", "record line 2 / MEDIAN EFF-DATE"),
    ("M_TYPE_CODE", "record line 2 / MEDIAN T Y"),
    ("M_CL", "record line 2 / MEDIAN C L"),
    ("M_BA", "record line 2 / MEDIAN B A"),
    ("M_WID+M_VA", "record line 2 / MEDIAN V WDA composite"),
    ("RT_SIG", "record line 2 / RIGHT ROADBED change flag"),
    ("R_EFF_DATE", "record line 2 / RIGHT ROADBED EFF-DATE"),
    ("R_ST", "record line 2 / RIGHT ROADBED S T"),
    ("R_NO_LANES", "record line 2 / RIGHT ROADBED NO LN"),
    ("R_SF", "record line 2 / RIGHT ROADBED S F"),
    ("R_IN_TOT", "record line 2 / RIGHT ROADBED IN-SH TO (3-character cell)"),
    ("R_IN_TR", "record line 2 / RIGHT ROADBED IN-SH TR (3-character cell)"),
    ("R_TR_WID", "record line 2 / RIGHT ROADBED T-W WID"),
    ("R_OT_TOT", "record line 2 / RIGHT ROADBED OT-SH TO (3-character cell)"),
    ("R_OT_TR", "record line 2 / RIGHT ROADBED OT-SH TR (3-character cell)"),
    ("REFERENCE_DATE", "report-parameter page / REFERENCE DATE"),
    ("EXTRACT_DATE", "report-parameter and data-page header / REPORT DATE"),
])

NOT_PRINTED_INTERNAL = collections.OrderedDict([
    ("THY_ID", "database surrogate identifier; no print column"),
    ("BREAK_DESC", "database break helper; the print has no separate column (LK-BACK is LK_BACK_ADT)"),
    ("SEG_ORDER_ID", "not printed as text; audited as within-DCR record order"),
])

DATE_TOKEN = r"(?:\d{2}-\d{2}-\d{2}|\+{1,8})"
NUM_TOKEN = r"(?:\d{1,3}|\+{1,3})"
DCR_RE = re.compile(
    r"DIST-CNTY-ROUTE\s+(?P<dist>\d{2})-(?P<county>[A-Z]{1,3})-"
    r"(?P<route>\d{3})(?:\s+(?P<route_suffix>[A-Z]))?"
)
L1_RE = re.compile(
    rf"^\s*(?:(?P<pp>[A-Z])\s+)?(?P<postmile>\d{{3}}\.\d{{3}})"
    rf"(?P<e_ind>[A-Z])?\s+(?P<length>\d{{3}}\.\d{{3}})\s+"
    rf"(?P<rec_date>{DATE_TOKEN})\s+(?P<hg>\S)\s+(?P<ac>\S)\s+"
    rf"(?P<acc_sig>[*Y])?(?P<acc_eff_date>{DATE_TOKEN})\s+"
    rf"(?:(?P<city>[A-Z][A-Z0-9-]{{0,4}})\s+)?(?P<pop_code>[RUB])\s+"
    rf"(?P<beg_date>{DATE_TOKEN})\s+(?P<adt_tail>.+?)\s*$"
)
L2_RE = re.compile(
    rf"^(?P<description>.*?)\s*(?P<non_add>[AN])\s+"
    rf"(?P<lt_sig>[*Y])?(?P<l_eff_date>{DATE_TOKEN})\s+"
    rf"(?P<l_st>\S+)\s+(?P<l_no_lanes>{NUM_TOKEN})\s+(?P<l_sf>\S+)\s+"
    rf"(?P<l_ot_tot>{NUM_TOKEN})\s+(?P<l_ot_tr>{NUM_TOKEN})\s+"
    rf"(?P<l_tr_wid>{NUM_TOKEN})\s+(?P<l_in_tot>{NUM_TOKEN})\s+"
    rf"(?P<l_in_tr>{NUM_TOKEN})\s+"
    rf"(?P<med_sig>[*Y])?(?P<m_eff_date>{DATE_TOKEN})\s+"
    rf"(?P<m_type_code>\S+)\s+(?P<m_cl>\S+)\s+(?P<m_ba>\S+)\s+"
    rf"(?P<m_wid_va>(?:\d{{1,3}}[A-Z]?|\+{{1,4}}))\s+"
    rf"(?P<rt_sig>[*Y])?(?P<r_eff_date>{DATE_TOKEN})\s+"
    rf"(?P<r_st>\S+)\s+(?P<r_no_lanes>{NUM_TOKEN})\s+(?P<r_sf>\S+)\s+"
    rf"(?P<r_in_tot>{NUM_TOKEN})\s+(?P<r_in_tr>{NUM_TOKEN})\s+"
    rf"(?P<r_tr_wid>{NUM_TOKEN})\s+(?P<r_ot_tot>{NUM_TOKEN})\s+"
    rf"(?P<r_ot_tr>{NUM_TOKEN})\s*$"
)
IDENTITY_L1_RE = re.compile(
    rf"^\s*(?:(?P<pp>[A-Z])\s+)?(?P<postmile>\d{{3}}\.\d{{3}})"
    rf"(?P<e_ind>[A-Z])?\s+(?P<length>\d{{3}}\.\d{{3}})\s+"
    rf"(?P<rec_date>{DATE_TOKEN})\s+(?P<hg>\S)(?:\s+|$)"
)
PM_CANDIDATE_RE = re.compile(r"^\s*(?:[A-Z]\s+)?\d{3}\.\d{3}[A-Z]?\s+")
REPORT_DATE_RE = re.compile(r"REPORT DATE\s*:\s*(\d{2}/\d{2}/\d{4})")
REFERENCE_DATE_RE = re.compile(r"REFERENCE DATE\s*:\s*(\d{2}/\d{2}/\d{4})")
PARAM_DISTRICT_RE = re.compile(r"DISTRICT\s+(\d{2})")
DATA_DATE_RE = re.compile(r"(\d{2}/\d{2}/\d{4})\s+TSAR-HIGHWAY DETAIL")

DIRECT_GROUPS = {
    "PP": "pp", "POSTMILE": "postmile", "E_IND": "e_ind",
    "LENGTH": "length", "REC_DATE": "rec_date", "HG": "hg", "AC": "ac",
    "ACC_SIG": "acc_sig", "ACC_EFF_DATE": "acc_eff_date", "CITY": "city",
    "POP_CODE": "pop_code", "BEG_DATE": "beg_date",
    "DESCRIPTION": "description", "NON_ADD": "non_add", "LT_SIG": "lt_sig",
    "L_EFF_DATE": "l_eff_date", "L_ST": "l_st", "L_NO_LANES": "l_no_lanes",
    "L_SF": "l_sf", "L_OT_TOT": "l_ot_tot", "L_OT_TR": "l_ot_tr",
    "L_TR_WID": "l_tr_wid", "L_IN_TOT": "l_in_tot", "L_IN_TR": "l_in_tr",
    "MED_SIG": "med_sig", "M_EFF_DATE": "m_eff_date",
    "M_TYPE_CODE": "m_type_code", "M_CL": "m_cl", "M_BA": "m_ba",
    "RT_SIG": "rt_sig", "R_EFF_DATE": "r_eff_date", "R_ST": "r_st",
    "R_NO_LANES": "r_no_lanes", "R_SF": "r_sf", "R_IN_TOT": "r_in_tot",
    "R_IN_TR": "r_in_tr", "R_TR_WID": "r_tr_wid", "R_OT_TOT": "r_ot_tot",
    "R_OT_TR": "r_ot_tr",
}

NUMERIC_RENDER_FIELDS = {
    "L_NO_LANES", "L_OT_TOT", "L_OT_TR", "L_TR_WID", "L_IN_TOT", "L_IN_TR",
    "R_NO_LANES", "R_IN_TOT", "R_IN_TR", "R_TR_WID", "R_OT_TOT", "R_OT_TR",
}
SHOULDER_PRINT_FIELDS = {
    "L_OT_TOT", "L_OT_TR", "L_IN_TOT", "L_IN_TR",
    "R_OT_TOT", "R_OT_TR", "R_IN_TOT", "R_IN_TR",
}


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _collapse(value) -> str:
    return " ".join(_text(value).split())


def _route_token(row: dict) -> str:
    route = _text(row["RTE"])
    route = f"{int(route):03d}" if route.isdigit() else route
    return f"{route}{_text(row['RTE_SFX'])}".upper()


def _mile(value) -> str:
    try:
        return f"{float(_text(value)):07.3f}"
    except ValueError:
        return _text(value)


def _roadbed_claim(value) -> str:
    value = _text(value).upper()
    return value if value in {"R", "L"} else ""


def _physical_key_from_xlsx(row: dict) -> tuple:
    return (
        _text(row["PP"]).upper(), _mile(row["POSTMILE"]),
        _text(row["E_IND"]).upper(), _roadbed_claim(row["HG"]),
    )


def _physical_key_from_pdf(record: dict) -> tuple:
    f = record["fields"]
    return (
        _text(f.get("PP")).upper(), _mile(f.get("POSTMILE")),
        _text(f.get("E_IND")).upper(), _roadbed_claim(f.get("HG")),
    )


def _full_identity(group: tuple, key: tuple) -> dict:
    dist, county, route = group
    pp, postmile, equation, roadbed = key
    base = route[:3]
    suffix = route[3:]
    return {
        "district": dist, "county": county, "route_base": base,
        "route_suffix": suffix, "pp": pp, "postmile": postmile,
        "equation_claim": equation, "roadbed_claim": roadbed,
    }


def _with_occurrence(identity: dict, occurrence: int | None) -> dict:
    out = dict(identity)
    if occurrence is not None:
        out["occurrence_within_physical_identity"] = occurrence
    return out


def _render_xlsx(field: str, row: dict) -> str:
    if field == "POSTMILE":
        return _mile(row[field])
    if field == "LENGTH":
        return _mile(row[field])
    if field == "DESCRIPTION":
        # The legacy print allocates exactly 23 source characters to DESCRIPTION;
        # fixed-width whitespace is layout and then collapses in extraction.
        return _collapse(_text(row[field])[:23])
    if field == "BEG_DATE":
        raw = _text(row[field])
        if re.fullmatch(r"\d{2}-\d{2}-\d{2}", raw):
            return f"{raw[:2]}-01-01"
        return raw
    if field in NUMERIC_RENDER_FIELDS:
        raw = _text(row[field])
        rendered = str(int(raw)) if raw.isdigit() else raw
        # Each shoulder TO/TR print cell is three characters wide.  Values such
        # as 12.8 visibly print as ``12.``; 6.5 and integral values fit.
        return rendered[:3] if field in SHOULDER_PRINT_FIELDS else rendered
    if field == "M_WID+M_VA":
        width, variance = _text(row["M_WID"]), _text(row["M_VA"]).upper()
        if width.isdigit():
            width = f"{int(width):02d}"
        return f"{width}{variance}"
    return _text(row.get(field))


def _raw_xlsx_display(field: str, row: dict) -> str:
    if field == "M_WID+M_VA":
        return f"{_text(row['M_WID'])}|{_text(row['M_VA'])}"
    return _text(row.get(field))


def _parse_adt_tail(tail: str):
    """Return the six ADT print cells, or a precise ambiguity reason.

    LK-AHD and PROFILE are the first two cells; CHANGE/MILE and DVM are the
    final two.  The middle zero-to-two tokens are BREAK_DESC (alphabetic marker)
    and/or LK_BACK_ADT (numeric).  This rule comes from the printed headings and
    is deliberately stricter than accepting an arbitrary token sequence.
    """
    tokens = tail.split()
    if len(tokens) < 4:
        return None, f"ADT tail has {len(tokens)} tokens, expected at least four: {tail!r}"
    adt, profile = tokens[0], tokens[1]
    middle, chngmile, dvm = tokens[2:-2], tokens[-2], tokens[-1]
    if profile not in {"P", "S"}:
        return None, f"ADT profile is not P/S: {tail!r}"
    break_desc = lk_back = ""
    if len(middle) == 1:
        if re.search(r"[A-Z-]", middle[0]):
            break_desc = middle[0]
        else:
            lk_back = middle[0]
    elif len(middle) == 2:
        break_desc, lk_back = middle
    elif len(middle) > 2:
        return None, f"ADT tail middle has {len(middle)} tokens: {tail!r}"
    return {
        "ADT_AMT": adt, "PROFILE": profile, "BREAK_DESC": break_desc,
        "LK_BACK_ADT": lk_back, "CHNGMILE": chngmile, "DVM": dvm,
    }, None


def _parse_record_lines(line1: str, line2: str):
    m1, m2 = L1_RE.match(line1), L2_RE.match(line2)
    if not m1:
        return None, f"line 1 did not match grammar: {line1!r}"
    if not m2:
        return None, f"line 2 did not match grammar: {line2!r}"
    fields = {}
    for field, group in DIRECT_GROUPS.items():
        match = m1 if group in m1.groupdict() else m2
        fields[field] = _text(match.group(group) or "")
    adt, error = _parse_adt_tail(m1.group("adt_tail"))
    if error:
        return None, error
    fields.update(adt)
    fields["M_WID+M_VA"] = _text(m2.group("m_wid_va"))
    return fields, None


LINE1_FIELDS = (
    "PP", "POSTMILE", "E_IND", "LENGTH", "REC_DATE", "HG", "AC", "ACC_SIG",
    "ACC_EFF_DATE", "CITY", "POP_CODE", "BEG_DATE", "ADT_AMT", "PROFILE",
    "LK_BACK_ADT", "CHNGMILE", "DVM",
)
LINE2_FIELDS = (
    "DESCRIPTION", "NON_ADD", "LT_SIG", "L_EFF_DATE", "L_ST", "L_NO_LANES",
    "L_SF", "L_OT_TOT", "L_OT_TR", "L_TR_WID", "L_IN_TOT", "L_IN_TR",
    "MED_SIG", "M_EFF_DATE", "M_TYPE_CODE", "M_CL", "M_BA", "M_WID+M_VA",
    "RT_SIG", "R_EFF_DATE", "R_ST", "R_NO_LANES", "R_SF", "R_IN_TOT",
    "R_IN_TR", "R_TR_WID", "R_OT_TOT", "R_OT_TR",
)

LINE1_TOKEN_COMPONENTS = (
    ("PP",), ("POSTMILE", "E_IND"), ("LENGTH",), ("REC_DATE",), ("HG",),
    ("AC",), ("ACC_SIG", "ACC_EFF_DATE"), ("CITY",), ("POP_CODE",),
    ("BEG_DATE",), ("ADT_AMT",), ("PROFILE",), ("LK_BACK_ADT",),
    ("CHNGMILE",), ("DVM",),
)


def _print_stream(value: str) -> str:
    """The vendor print is fixed-position text; whitespace is layout, not data."""
    return re.sub(r"\s+", "", value or "")


def _expected_components(row: dict, fields: tuple[str, ...]):
    return [(field, _print_stream(_render_xlsx(field, row))) for field in fields]


def _stream_diff(row: dict, pdf_line: str, fields: tuple[str, ...]) -> dict:
    """Compare one physical print line without trusting PDF word separators.

    Some real pages have adjacent numeric cells extracted as one token (for
    example 24 | 6.5 | 6.5 becomes ``246.56.5``).  Comparing the fixed-order
    character stream proves the printed cells without pretending those missing
    extraction spaces are data differences.  SequenceMatcher is used only to
    attribute a non-equal stream back to XLSX field spans; a boundary insertion
    is explicitly unsafe and is returned as unresolved attribution.
    """
    components = _expected_components(row, fields)
    expected = "".join(value for _field, value in components)
    actual = _print_stream(pdf_line)
    offsets = []
    cursor = 0
    for field, value in components:
        offsets.append((field, cursor, cursor + len(value), value))
        cursor += len(value)
    if expected == actual:
        return {
            "equal": True, "changed_fields": [], "unsafe_attribution": [],
            "expected_stream": expected, "actual_stream": actual, "opcodes": [],
        }

    changed = set()
    unsafe = []
    opcodes = []
    matcher = difflib.SequenceMatcher(None, expected, actual, autojunk=False)
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        touched = [field for field, start, end, _value in offsets
                   if i1 < end and i2 > start]
        if tag == "insert":
            inside = [field for field, start, end, _value in offsets if start < i1 < end]
            if inside:
                touched = inside
            else:
                left = [field for field, start, end, _value in offsets if end == i1]
                right = [field for field, start, end, _value in offsets if start == i1]
                touched = list(dict.fromkeys(left[-1:] + right[:1]))
                if len(touched) != 1:
                    unsafe.append({
                        "kind": "insertion_at_field_boundary", "expected_offset": i1,
                        "candidate_fields": touched, "inserted": actual[j1:j2],
                    })
        if not touched:
            unsafe.append({
                "kind": "difference_not_attributable_to_nonblank_field",
                "expected_span": [i1, i2], "actual_span": [j1, j2],
            })
        changed.update(touched)
        opcodes.append({
            "tag": tag, "expected_span": [i1, i2], "actual_span": [j1, j2],
            "expected": expected[i1:i2], "actual": actual[j1:j2],
            "candidate_fields": touched,
        })
    return {
        "equal": False, "changed_fields": sorted(changed),
        "unsafe_attribution": unsafe, "expected_stream": expected,
        "actual_stream": actual, "opcodes": opcodes,
    }


def _line1_token_diff(row: dict, pdf_line: str) -> dict:
    """Field-safe line-one diff using the print's whitespace-delimited cells."""
    all_components = []
    for fields in LINE1_TOKEN_COMPONENTS:
        value = "".join(_render_xlsx(field, row) for field in fields)
        all_components.append((fields, value))
    present = [(original_i, fields, value) for original_i, (fields, value)
               in enumerate(all_components) if value]
    expected = [value for _i, _fields, value in present]
    actual = pdf_line.split()
    if expected == actual:
        return {
            "equal": True, "changed_fields": [], "unsafe_attribution": [],
            "expected_stream": expected, "actual_stream": actual, "opcodes": [],
        }
    changed = set()
    unsafe = []
    opcodes = []
    matcher = difflib.SequenceMatcher(None, expected, actual, autojunk=False)
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        touched = []
        for _original_i, fields, _value in present[i1:i2]:
            touched.extend(fields)
        if tag == "insert":
            left_original = present[i1 - 1][0] if i1 else -1
            right_original = present[i1][0] if i1 < len(present) else len(all_components)
            blank_between = [
                fields for original_i, (fields, value) in enumerate(all_components)
                if left_original < original_i < right_original and not value
            ]
            if len(blank_between) == 1:
                touched = list(blank_between[0])
            else:
                neighbors = []
                if i1:
                    neighbors.extend(present[i1 - 1][1])
                if i1 < len(present):
                    neighbors.extend(present[i1][1])
                touched = list(dict.fromkeys(neighbors))
                unsafe.append({
                    "kind": "inserted_token_has_no_unique_blank_print_field",
                    "expected_token_offset": i1,
                    "candidate_fields": touched,
                    "inserted": actual[j1:j2],
                })
        changed.update(touched)
        opcodes.append({
            "tag": tag, "expected_span": [i1, i2], "actual_span": [j1, j2],
            "expected": expected[i1:i2], "actual": actual[j1:j2],
            "candidate_fields": list(dict.fromkeys(touched)),
        })
    return {
        "equal": False, "changed_fields": sorted(changed),
        "unsafe_attribution": unsafe, "expected_stream": expected,
        "actual_stream": actual, "opcodes": opcodes,
    }


def _date_iso(mmddyyyy: str) -> str:
    return dt.datetime.strptime(mmddyyyy, "%m/%d/%Y").date().isoformat()


def read_xlsx(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if wb.sheetnames != ["Sheet 1"]:
            raise ValueError(f"workbook sheets are {wb.sheetnames!r}, expected ['Sheet 1']")
        ws = wb["Sheet 1"]
        iterator = ws.iter_rows(values_only=True)
        header = tuple(next(iterator, ()) or ())
        if header != EXPECTED_XLSX_HEADER:
            raise ValueError("workbook 56-column header is not the frozen authoritative schema")
        rows = []
        for source_row, values in enumerate(iterator, 2):
            if not any(v not in (None, "") for v in values):
                continue
            row = dict(zip(header, values))
            row["__source_row__"] = source_row
            rows.append(row)
    finally:
        wb.close()

    refs = sorted({_text(r["REFERENCE_DATE"]) for r in rows})
    extracts = sorted({_text(r["EXTRACT_DATE"]) for r in rows})
    groups = collections.defaultdict(list)
    for row in rows:
        groups[(_text(row["DIST"]), _text(row["CNTY"]), _route_token(row))].append(row)
    for group_rows in groups.values():
        group_rows.sort(key=lambda r: (
            float(r["SEG_ORDER_ID"]) if r["SEG_ORDER_ID"] is not None else float("inf"),
            r["__source_row__"],
        ))

    duplicate_groups = []
    ambiguous = []
    for group, group_rows in groups.items():
        by_key = collections.defaultdict(list)
        for row in group_rows:
            by_key[_physical_key_from_xlsx(row)].append(row)
        for key, copies in by_key.items():
            if len(copies) <= 1:
                continue
            signatures = {
                tuple(_render_xlsx(field, row) for field in FIELD_TO_PDF
                      if field not in {"REFERENCE_DATE", "EXTRACT_DATE"})
                for row in copies
            }
            item = {
                "identity": _full_identity(group, key), "multiplicity": len(copies),
                "distinct_printable_signatures": len(signatures),
                "xlsx_rows": [r["__source_row__"] for r in copies[:20]],
            }
            duplicate_groups.append(item)
            if len(signatures) > 1:
                ambiguous.append(item)

    return {
        "rows": rows, "groups": groups, "reference_dates": refs,
        "extract_dates": extracts, "duplicate_groups": duplicate_groups,
        "ambiguous_duplicate_groups": ambiguous,
    }


def inspect_pdf_metadata(path: Path, expected_dist: str) -> dict:
    with pdfplumber.open(path) as pdf:
        pages = len(pdf.pages)
        parameter_text = "\n".join(
            (pdf.pages[i].extract_text(x_tolerance=2, y_tolerance=3) or "")
            for i in range(min(3, pages))
        )
    report = REPORT_DATE_RE.search(parameter_text)
    reference = REFERENCE_DATE_RE.search(parameter_text)
    district = PARAM_DISTRICT_RE.search(parameter_text)
    digest = _sha256(path)
    return {
        "member": path.name,
        "bytes": path.stat().st_size,
        "sha256": digest,
        "sha256_expected": EXPECTED_PDF_SHA256.get(path.name),
        "sha256_matches": digest == EXPECTED_PDF_SHA256.get(path.name),
        "pages": pages,
        "filename_district": expected_dist,
        "internal_district": district.group(1) if district else None,
        "report_date": _date_iso(report.group(1)) if report else None,
        "reference_date": _date_iso(reference.group(1)) if reference else None,
        "parameter_page_valid": bool(
            report and reference and district and district.group(1) == expected_dist
            and "TSAR - HIGHWAY DETAIL" in parameter_text
        ),
    }


def parse_pdf(path: Path, expected_dist: str, progress_every: int = 100) -> dict:
    records = []
    residue = []
    dcr = None
    pending = None
    data_pages = 0
    data_header_dates = collections.Counter()
    dcr_headers = collections.Counter()
    started = time.monotonic()
    with pdfplumber.open(path) as pdf:
        for page_index, page in enumerate(pdf.pages, 1):
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            page_dates = DATA_DATE_RE.findall(text)
            if page_dates:
                data_pages += 1
                data_header_dates.update(_date_iso(d) for d in page_dates)
            lines = [_collapse(line) for line in text.splitlines() if _collapse(line)]
            for line_index, line in enumerate(lines, 1):
                dcr_match = DCR_RE.search(line)
                if dcr_match:
                    route = dcr_match.group("route") + (dcr_match.group("route_suffix") or "")
                    dcr = (dcr_match.group("dist"), dcr_match.group("county"), route)
                    dcr_headers[dcr] += 1
                    if dcr[0] != expected_dist:
                        residue.append({
                            "kind": "wrong_internal_district", "page": page_index,
                            "line": line_index, "text": line, "dcr": dcr,
                        })
                    pending = None
                    continue
                if dcr is None:
                    continue
                m1 = IDENTITY_L1_RE.match(line)
                if m1:
                    if pending is not None:
                        residue.append({
                            "kind": "line1_replaced_before_line2", "page": page_index,
                            "line": line_index, "previous": pending["line"], "text": line,
                            "dcr": dcr,
                        })
                    pending = {
                        "line": line, "page": page_index, "line_number": line_index,
                        "dcr": dcr,
                    }
                    continue
                if pending is not None:
                    if not IDENTITY_L1_RE.match(line):
                        identity_match = IDENTITY_L1_RE.match(pending["line"])
                        fields = {
                            "PP": _text(identity_match.group("pp") or ""),
                            "POSTMILE": _text(identity_match.group("postmile")),
                            "E_IND": _text(identity_match.group("e_ind") or ""),
                            "HG": _text(identity_match.group("hg")),
                        }
                        records.append({
                            "dcr": pending["dcr"], "fields": fields,
                            "line1": pending["line"], "line2": line,
                            "page_line1": pending["page"],
                            "line_number1": pending["line_number"],
                            "page_line2": page_index, "line_number2": line_index,
                        })
                        pending = None
                        continue
                if PM_CANDIDATE_RE.match(line):
                    residue.append({
                        "kind": "postmile_candidate_not_line1", "page": page_index,
                        "line": line_index, "text": line, "dcr": dcr,
                    })
            if progress_every and page_index % progress_every == 0:
                elapsed = time.monotonic() - started
                print(
                    f"{path.name}: {page_index:,}/{len(pdf.pages):,} pages, "
                    f"{len(records):,} records, {len(residue):,} residue ({elapsed:.1f}s)",
                    flush=True,
                )
    if pending is not None:
        residue.append({
            "kind": "document_ended_with_line1", "page": pending["page"],
            "line": pending["line_number"], "text": pending["line"],
            "dcr": pending["dcr"],
        })
    return {
        "records": records, "residue": residue, "data_pages": data_pages,
        "data_header_dates": dict(sorted(data_header_dates.items())),
        "dcr_headers": dcr_headers,
    }


def _sample_append(samples: list, item: dict, limit: int = 30):
    if len(samples) < limit:
        samples.append(item)


def _edit_cost(expected: str, actual: str) -> int:
    matcher = difflib.SequenceMatcher(None, expected, actual, autojunk=False)
    return sum(max(i2 - i1, j2 - j1)
               for tag, i1, i2, j1, j2 in matcher.get_opcodes() if tag != "equal")


def _row_record_cost(row: dict, rec: dict) -> int:
    expected_1 = "".join(value for _field, value in _expected_components(row, LINE1_FIELDS))
    expected_2 = "".join(value for _field, value in _expected_components(row, LINE2_FIELDS))
    return (_edit_cost(expected_1, _print_stream(rec["line1"]))
            + _edit_cost(expected_2, _print_stream(rec["line2"])))


def reconcile(xlsx: dict, pdf_records: list, date_skew_is_bound: bool) -> dict:
    selected_districts = {rec["dcr"][0] for rec in pdf_records}
    raw_by_full_key = collections.defaultdict(list)
    for group, rows in xlsx["groups"].items():
        if group[0] not in selected_districts:
            continue
        occurrences = collections.Counter()
        for row in rows:
            key = _physical_key_from_xlsx(row)
            occurrences[key] += 1
            row["__physical_occurrence__"] = occurrences[key]
            raw_by_full_key[(group, key)].append(row)

    paired = []
    pdf_only = []
    pdf_duplicate_groups = []
    pdf_by_full_key = collections.defaultdict(list)
    pdf_occurrences = collections.Counter()
    for rec in pdf_records:
        full_key = (rec["dcr"], _physical_key_from_pdf(rec))
        pdf_occurrences[full_key] += 1
        rec["physical_occurrence"] = pdf_occurrences[full_key]
        pdf_by_full_key[full_key].append(rec)
    for (group, key), copies in pdf_by_full_key.items():
        if len(copies) > 1:
            signatures = {(c["line1"], c["line2"]) for c in copies}
            pdf_duplicate_groups.append({
                "identity": _full_identity(group, key), "multiplicity": len(copies),
                "distinct_printable_signatures": len(signatures),
                "pages": [c["page_line1"] for c in copies[:20]],
            })

    xlsx_only = []
    pairing_decisions = []
    for full_key in sorted(set(raw_by_full_key) | set(pdf_by_full_key)):
        group, key = full_key
        raw_rows = raw_by_full_key.get(full_key, [])
        printed = pdf_by_full_key.get(full_key, [])
        nr, np = len(raw_rows), len(printed)
        if not nr:
            pdf_only.extend(printed)
            continue
        if not np:
            xlsx_only.extend((group, key, row) for row in raw_rows)
            continue
        if nr >= np:
            choices = itertools.permutations(range(nr), np)
            best = min(
                choices,
                key=lambda selected: (
                    sum(_row_record_cost(raw_rows[ri], printed[pi])
                        for pi, ri in enumerate(selected)),
                    selected,
                ),
            )
            selected_raw = set(best)
            paired.extend((raw_rows[ri], printed[pi]) for pi, ri in enumerate(best))
            xlsx_only.extend((group, key, row) for ri, row in enumerate(raw_rows)
                             if ri not in selected_raw)
            chosen_raw, chosen_pdf = list(best), list(range(np))
        else:
            choices = itertools.permutations(range(np), nr)
            best = min(
                choices,
                key=lambda selected: (
                    sum(_row_record_cost(raw_rows[ri], printed[pi])
                        for ri, pi in enumerate(selected)),
                    selected,
                ),
            )
            selected_pdf = set(best)
            paired.extend((raw_rows[ri], printed[pi]) for ri, pi in enumerate(best))
            pdf_only.extend(rec for pi, rec in enumerate(printed) if pi not in selected_pdf)
            chosen_raw, chosen_pdf = list(range(nr)), list(best)
        if nr > 1 or np > 1:
            pairing_decisions.append({
                "identity": _full_identity(group, key),
                "xlsx_multiplicity": nr, "pdf_multiplicity": np,
                "chosen_xlsx_occurrences": [i + 1 for i in chosen_raw],
                "chosen_pdf_occurrences": [i + 1 for i in chosen_pdf],
                "pair_cost": sum(
                    _row_record_cost(raw_rows[ri], printed[pi])
                    for ri, pi in zip(chosen_raw, chosen_pdf)
                ),
            })

    exact_cells = 0
    render_equivalent_cells = 0
    source_date_delta_cells = 0
    unresolved_cells = 0
    cell_totals = collections.Counter()
    render_rules = collections.Counter()
    delta_fields = collections.Counter()
    unresolved_fields = collections.Counter()
    delta_samples = []
    unresolved_samples = []
    source_delta_line_count = 0
    unsafe_stream_differences = []
    delta_manifest_items = []

    for row, rec in paired:
        group, key = rec["dcr"], _physical_key_from_pdf(rec)
        occurrence = rec["physical_occurrence"]
        identity = _with_occurrence(_full_identity(group, key), occurrence)
        route_pdf = {
            "DIST": group[0], "CNTY": group[1], "RTE": group[2][:3],
            "RTE_SFX": group[2][3:], "DIST_CNTY_ROUTE":
                f"{group[0]}-{group[1]}-{group[2][:3]}" +
                (f" {group[2][3:]}" if group[2][3:] else ""),
        }
        for field in ("DIST", "CNTY", "RTE", "RTE_SFX", "DIST_CNTY_ROUTE"):
            expected = _render_xlsx(field, row)
            actual = _text(route_pdf[field])
            cell_totals[field] += 1
            if expected == actual:
                exact_cells += 1
                continue
            item = {
                "identity": identity, "field": field,
                "xlsx": expected, "pdf": actual, "xlsx_row": row["__source_row__"],
                "pdf_page": rec["page_line1"],
            }
            if date_skew_is_bound:
                source_date_delta_cells += 1
                delta_fields[field] += 1
                _sample_append(delta_samples, item)
                delta_manifest_items.append({
                    "kind": "route_header_field", "identity": identity,
                    "field": field, "xlsx": expected, "pdf": actual,
                })
            else:
                unresolved_cells += 1
                unresolved_fields[field] += 1
                _sample_append(unresolved_samples, item)

        for fields, pdf_line, line_number in (
            (LINE1_FIELDS, rec["line1"], 1),
            (LINE2_FIELDS, rec["line2"], 2),
        ):
            stream = (_line1_token_diff(row, pdf_line) if line_number == 1
                      else _stream_diff(row, pdf_line, fields))
            changed = set(stream["changed_fields"])
            for field in fields:
                cell_totals[field] += 1
                raw_display = _raw_xlsx_display(field, row)
                rendered = _render_xlsx(field, row)
                if raw_display != rendered:
                    render_equivalent_cells += 1
                    if field in {"POSTMILE", "LENGTH"}:
                        render_rules["fixed three-decimal print formatting"] += 1
                    elif field == "DESCRIPTION":
                        render_rules["DESCRIPTION first 23 source characters plus whitespace collapse"] += 1
                    elif field == "BEG_DATE":
                        render_rules["ADT begin date prints as count-year YY-01-01"] += 1
                    elif field in NUMERIC_RENDER_FIELDS:
                        if field in SHOULDER_PRINT_FIELDS:
                            render_rules["three-character shoulder TO/TR print cells"] += 1
                        else:
                            render_rules["numeric de-padding"] += 1
                    elif field == "M_WID+M_VA":
                        render_rules["median width plus variance composition"] += 1
                if field not in changed:
                    exact_cells += 1
                    continue
                if date_skew_is_bound:
                    source_date_delta_cells += 1
                    delta_fields[field] += 1
                else:
                    unresolved_cells += 1
                    unresolved_fields[field] += 1
            if not stream["equal"]:
                source_delta_line_count += 1
                sample = {
                    "identity": identity, "pdf_line": line_number,
                    "xlsx_row": row["__source_row__"],
                    "pdf_page": rec[f"page_line{line_number}"],
                    "changed_fields": stream["changed_fields"],
                    "opcodes": stream["opcodes"],
                }
                if date_skew_is_bound:
                    _sample_append(delta_samples, sample)
                    delta_manifest_items.append({
                        "kind": "print_line", "identity": identity,
                        "pdf_line": line_number,
                        "changed_fields": stream["changed_fields"],
                        "xlsx_stream": stream["expected_stream"],
                        "pdf_stream": stream["actual_stream"],
                    })
                else:
                    _sample_append(unresolved_samples, sample)
                if stream["unsafe_attribution"]:
                    _sample_append(unsafe_stream_differences, {
                        **sample, "unsafe_attribution": stream["unsafe_attribution"],
                    }, 100)

    # SEG_ORDER_ID is not printed.  Its auditable representation is that paired
    # XLSX records occur monotonically within each DCR in PDF sequence.
    paired_by_group = collections.defaultdict(list)
    for row, rec in paired:
        paired_by_group[rec["dcr"]].append((row, rec))
    order_inversions = []
    for group, pairs in paired_by_group.items():
        pairs.sort(key=lambda pair: (
            pair[1]["page_line1"], pair[1]["line_number1"],
            pair[1]["page_line2"], pair[1]["line_number2"],
        ))
        previous = None
        for row, rec in pairs:
            current = float(row["SEG_ORDER_ID"])
            if previous is not None and current < previous:
                _sample_append(order_inversions, {
                    "dcr": group, "previous_seg_order_id": previous,
                    "seg_order_id": current, "xlsx_row": row["__source_row__"],
                    "pdf_page": rec["page_line1"],
                })
            previous = current

    for group, key, row in xlsx_only:
        delta_manifest_items.append({
            "kind": "xlsx_only_record",
            "identity": _with_occurrence(
                _full_identity(group, key), row.get("__physical_occurrence__")),
            "line1_stream": "".join(
                value for _field, value in _expected_components(row, LINE1_FIELDS)),
            "line2_stream": "".join(
                value for _field, value in _expected_components(row, LINE2_FIELDS)),
        })
    for rec in pdf_only:
        delta_manifest_items.append({
            "kind": "pdf_only_record",
            "identity": _with_occurrence(
                _full_identity(rec["dcr"], _physical_key_from_pdf(rec)),
                rec.get("physical_occurrence")),
            "line1_stream": _print_stream(rec["line1"]),
            "line2_stream": _print_stream(rec["line2"]),
        })
    delta_manifest_items.sort(
        key=lambda item: json.dumps(item, sort_keys=True, separators=(",", ":")))
    manifest_payload = "\n".join(
        json.dumps(item, sort_keys=True, separators=(",", ":"))
        for item in delta_manifest_items
    )
    delta_manifest_sha256 = hashlib.sha256(
        manifest_payload.encode("utf-8")
    ).hexdigest()

    return {
        "pdf_records": len(pdf_records), "xlsx_records": len(xlsx["rows"]),
        "xlsx_records_in_pdf_scope": sum(
            len(rows) for group, rows in xlsx["groups"].items()
            if group[0] in selected_districts
        ),
        "paired_records": len(paired), "xlsx_only_records": len(xlsx_only),
        "pdf_only_records": len(pdf_only),
        "membership_delta_class": (
            "observed_cross_snapshot_delta_pending_exact_allowlist"
            if date_skew_is_bound and (xlsx_only or pdf_only)
            else "none" if not (xlsx_only or pdf_only) else "unresolved"
        ),
        "xlsx_only_samples": [
            {"identity": _with_occurrence(_full_identity(g, k), r.get("__physical_occurrence__")),
             "xlsx_row": r["__source_row__"]}
            for g, k, r in xlsx_only[:30]
        ],
        "pdf_only_samples": [
            {"identity": _with_occurrence(
                 _full_identity(r["dcr"], _physical_key_from_pdf(r)),
                 r.get("physical_occurrence")),
             "pdf_page": r["page_line1"]}
            for r in pdf_only[:30]
        ],
        "exact_printable_cells": exact_cells,
        "render_equivalent_cells": render_equivalent_cells,
        "render_equivalence_rules": dict(sorted(render_rules.items())),
        "source_date_delta_cells": source_date_delta_cells,
        "source_date_delta_by_field": dict(sorted(delta_fields.items())),
        "source_date_delta_line_count": source_delta_line_count,
        "source_date_delta_samples": delta_samples,
        "observed_delta_manifest": {
            "item_count": len(delta_manifest_items),
            "sha256": delta_manifest_sha256,
            "field_counts": dict(sorted(delta_fields.items())),
            "items": delta_manifest_items,
        },
        "unresolved_cells": unresolved_cells,
        "unresolved_by_field": dict(sorted(unresolved_fields.items())),
        "unresolved_samples": unresolved_samples,
        "unsafe_stream_attribution_count": len(unsafe_stream_differences),
        "unsafe_stream_attribution_samples": unsafe_stream_differences,
        "compared_cells_by_field": dict(sorted(cell_totals.items())),
        "pdf_duplicate_groups": pdf_duplicate_groups,
        "duplicate_pairing_policy": (
            "maximum-cardinality, minimum-character-difference assignment within each "
            "complete physical identity; occurrence order is allowed to permute when "
            "same-key SEG_ORDER_ID values do not distinguish the records"
        ),
        "duplicate_pairing_decisions": pairing_decisions,
        "seg_order_inversion_count": len(order_inversions),
        "seg_order_inversion_samples": order_inversions,
    }


def _pdf_paths(pdf_dir: Path, districts: list[str]) -> list[tuple[str, Path]]:
    out = []
    for district in districts:
        matches = sorted(pdf_dir.glob(f"D{district} Highway Detail_TSN.pdf"))
        if len(matches) != 1:
            raise ValueError(
                f"district {district} has {len(matches)} canonical PDF members, expected one"
            )
        out.append((district, matches[0]))
    return out


def _write_json(path: Path, result: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(result, indent=2, sort_keys=False), encoding="utf-8")
    temp.replace(path)


def _matches_delta_allowlist(sha256: str, item_count: int, field_counts: dict) -> bool:
    return bool(
        EXPECTED_DELTA_MANIFEST_SHA256
        and sha256 == EXPECTED_DELTA_MANIFEST_SHA256
        and item_count == EXPECTED_DELTA_ITEM_COUNT
        and field_counts == EXPECTED_DELTA_FIELD_COUNTS
    )


def _run_internal_negative_checks() -> dict:
    row = {name: None for name in EXPECTED_XLSX_HEADER}
    row.update({
        "DIST": "01", "CNTY": "MEN", "RTE": "001", "RTE_SFX": "S",
        "PP": "R", "POSTMILE": "1.2", "E_IND": "E", "LENGTH": 0.1,
        "REC_DATE": "64-01-01", "HG": "U", "AC": "C",
        "ACC_EFF_DATE": "64-01-01", "POP_CODE": "R", "BEG_DATE": "22-09-05",
        "ADT_AMT": "003000", "PROFILE": "P", "LK_BACK_ADT": "004400",
        "CHNGMILE": "00000001.00000", "DVM": "00000001.000",
        "DESCRIPTION": None, "NON_ADD": "A", "L_EFF_DATE": "64-01-01",
        "L_ST": "H", "L_NO_LANES": "1", "L_SF": "Z", "L_OT_TOT": "2",
        "L_OT_TR": "2", "L_TR_WID": "12", "L_IN_TOT": "0", "L_IN_TR": "0",
        "M_EFF_DATE": "64-01-01", "M_TYPE_CODE": "B", "M_CL": "7",
        "M_BA": "Z", "M_WID": 0, "M_VA": "Z", "R_EFF_DATE": "64-01-01",
        "R_ST": "H", "R_NO_LANES": "1", "R_SF": "Z", "R_IN_TOT": "0",
        "R_IN_TR": "0", "R_TR_WID": "12", "R_OT_TOT": "2", "R_OT_TR": "2",
    })
    line1_tokens = []
    for fields in LINE1_TOKEN_COMPONENTS:
        token = "".join(_render_xlsx(field, row) for field in fields)
        if token:
            line1_tokens.append(token)
    line1 = " ".join(line1_tokens)
    if not _line1_token_diff(row, line1)["equal"]:
        raise AssertionError("negative-check baseline line 1 is not equal")
    length_mutation = list(line1_tokens)
    length_mutation[2] = "000.101"
    length_diff = _line1_token_diff(row, " ".join(length_mutation))
    if length_diff["equal"] or "LENGTH" not in length_diff["changed_fields"]:
        raise AssertionError("arbitrary LENGTH mutation was not rejected")
    beg_mutation = list(line1_tokens)
    beg_i = next(i for i, token in enumerate(line1_tokens) if token == "22-01-01")
    beg_mutation[beg_i] = "23-01-01"
    beg_diff = _line1_token_diff(row, " ".join(beg_mutation))
    if beg_diff["equal"] or "BEG_DATE" not in beg_diff["changed_fields"]:
        raise AssertionError("BEG_DATE exception outside YY-01-01 was not rejected")
    line2 = "".join(value for _field, value in _expected_components(row, LINE2_FIELDS))
    description_diff = _stream_diff(row, "UNLISTED" + line2, LINE2_FIELDS)
    if description_diff["equal"] or description_diff["changed_fields"] != ["DESCRIPTION"]:
        raise AssertionError("arbitrary DESCRIPTION insertion was not isolated")
    identity = _full_identity(("01", "MEN", "001S"), ("R", "001.200", "E", ""))
    if identity["county"] != "MEN" or identity["route_suffix"] != "S":
        raise AssertionError("county/route-suffix identity claim was lost")
    if "BREAK_DESC" in FIELD_TO_PDF or "BREAK_DESC" not in NOT_PRINTED_INTERNAL:
        raise AssertionError("nonprinted BREAK_DESC leaked into the report field map")
    if not _matches_delta_allowlist(
        EXPECTED_DELTA_MANIFEST_SHA256,
        EXPECTED_DELTA_ITEM_COUNT,
        EXPECTED_DELTA_FIELD_COUNTS,
    ):
        raise AssertionError("exact frozen delta allowlist did not match itself")
    if any((
        _matches_delta_allowlist(
            "0" * 64, EXPECTED_DELTA_ITEM_COUNT, EXPECTED_DELTA_FIELD_COUNTS),
        _matches_delta_allowlist(
            EXPECTED_DELTA_MANIFEST_SHA256,
            EXPECTED_DELTA_ITEM_COUNT + 1,
            EXPECTED_DELTA_FIELD_COUNTS,
        ),
        _matches_delta_allowlist(
            EXPECTED_DELTA_MANIFEST_SHA256,
            EXPECTED_DELTA_ITEM_COUNT,
            {**EXPECTED_DELTA_FIELD_COUNTS, "ARBITRARY": 1},
        ),
    )):
        raise AssertionError("mutated delta allowlist was accepted")
    return {
        "status": "pass", "cases": 10,
        "mutations": [
            "arbitrary LENGTH value", "BEG_DATE outside deterministic year projection",
            "inserted DESCRIPTION", "county retention", "route-suffix retention",
            "BREAK_DESC nonprint classification", "baseline line conservation",
            "mutated delta digest", "mutated delta item count", "mutated delta field counts",
        ],
    }


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--pdf-dir", type=Path, required=True)
    parser.add_argument("--result", type=Path, required=True)
    parser.add_argument(
        "--district", action="append", choices=[f"{n:02d}" for n in range(1, 13)],
        help="pilot one or more districts; omit to require and parse all D01-D12",
    )
    parser.add_argument("--progress-every", type=int, default=100)
    parser.add_argument(
        "--discover-allowlist", action="store_true",
        help="full-run only: emit the exact observed delta manifest without accepting it",
    )
    args = parser.parse_args(argv)

    started = time.monotonic()
    districts = args.district or [f"{n:02d}" for n in range(1, 13)]
    full_run = not args.district
    result = {
        "oracle": "phase4_highway_detail_tsn_pdf_oracle",
        "oracle_version": 1,
        "independence": {
            "production_comparator_imported": False,
            "production_parser_imported": False,
            "production_evidence_imported": False,
            "production_constants_imported": False,
        },
        "scope": {
            "districts": districts, "full_d01_d12_run": full_run,
            "required_page_count_for_full_run": EXPECTED_PDF_PAGES,
        },
        "field_to_pdf_mapping": FIELD_TO_PDF,
        "not_printed_internal_fields": NOT_PRINTED_INTERNAL,
        "row_to_report_view_mapping": {
            "unit": "one XLSX row maps to one two-line PDF record under its DCR header",
            "line_1_sections": ["location/record/access/city", "ADT INFORMATION"],
            "line_2_sections": ["DESCRIPTION/NON-ADD", "LEFT ROADBED", "MEDIAN", "RIGHT ROADBED"],
            "ordering": "SEG_ORDER_ID is represented by monotonically increasing record order within DCR",
        },
        "internal_negative_checks": _run_internal_negative_checks(),
    }

    exit_code = 0
    try:
        xlsx_hash = _sha256(args.xlsx)
        print(f"reading XLSX {args.xlsx}", flush=True)
        xlsx = read_xlsx(args.xlsx)
        result["xlsx_source"] = {
            "path": str(args.xlsx), "bytes": args.xlsx.stat().st_size,
            "sha256": xlsx_hash, "sha256_expected": EXPECTED_XLSX_SHA256,
            "sha256_matches": xlsx_hash == EXPECTED_XLSX_SHA256,
            "sheet": "Sheet 1", "columns": len(EXPECTED_XLSX_HEADER),
            "rows": len(xlsx["rows"]), "expected_rows": EXPECTED_XLSX_ROWS,
            "reference_dates": xlsx["reference_dates"],
            "extract_dates": xlsx["extract_dates"],
            "physical_duplicate_groups": len(xlsx["duplicate_groups"]),
            "physical_duplicate_rows": sum(d["multiplicity"] for d in xlsx["duplicate_groups"]),
            "duplicate_resolution": (
                "multiplicity is preserved; the reported identity retains a 1-based source "
                "occurrence, while pairing is maximum-cardinality/minimum-difference and may "
                "permute records whose SEG_ORDER_ID is tied"
            ),
            "ambiguous_nonidentical_duplicate_groups": xlsx["ambiguous_duplicate_groups"],
        }

        members = _pdf_paths(args.pdf_dir, districts)
        metadata = []
        all_records = []
        all_residue = []
        all_data_dates = collections.Counter()
        dcr_headers = collections.Counter()
        for district, path in members:
            meta = inspect_pdf_metadata(path, district)
            metadata.append(meta)
            print(f"parsing {path.name} ({meta['pages']:,} pages)", flush=True)
            parsed = parse_pdf(path, district, progress_every=args.progress_every)
            all_records.extend(parsed["records"])
            for residue in parsed["residue"]:
                residue = dict(residue)
                residue["member"] = path.name
                _sample_append(all_residue, residue, 200)
            all_data_dates.update(parsed["data_header_dates"])
            dcr_headers.update(parsed["dcr_headers"])
            meta["parsed_records"] = len(parsed["records"])
            meta["parse_residue_count"] = len(parsed["residue"])
            meta["data_pages"] = parsed["data_pages"]
            meta["data_header_dates"] = parsed["data_header_dates"]

        total_pages = sum(m["pages"] for m in metadata)
        result["pdf_sources"] = {
            "directory": str(args.pdf_dir), "members": metadata,
            "member_count": len(metadata), "total_pages": total_pages,
            "expected_members_for_full_run": EXPECTED_PDF_MEMBERS,
            "expected_pages_for_full_run": EXPECTED_PDF_PAGES,
            "parsed_records": len(all_records),
            "parse_residue_count": sum(m["parse_residue_count"] for m in metadata),
            "parse_residue_samples": all_residue,
            "data_header_dates": dict(sorted(all_data_dates.items())),
            "distinct_dcr_headers": len(dcr_headers),
        }

        pdf_report_dates = sorted({m["report_date"] for m in metadata})
        pdf_reference_dates = sorted({m["reference_date"] for m in metadata})
        date_skew_is_bound = bool(
            xlsx["reference_dates"] == [EXPECTED_XLSX_REFERENCE_DATE]
            and xlsx["extract_dates"] == [EXPECTED_XLSX_EXTRACT_DATE]
            and pdf_report_dates == [EXPECTED_PDF_REPORT_DATE]
            and pdf_reference_dates == [EXPECTED_PDF_REFERENCE_DATE]
        )
        result["source_snapshot_relation"] = {
            "xlsx_reference_dates": xlsx["reference_dates"],
            "xlsx_extract_dates": xlsx["extract_dates"],
            "pdf_report_dates": pdf_report_dates,
            "pdf_reference_dates": pdf_reference_dates,
            "reference_delta_days": 7 if date_skew_is_bound else None,
            "exact_expected_date_skew_bound": date_skew_is_bound,
            "classification_rule": (
                "real printable-value or membership differences are exact source-snapshot "
                "deltas only while all four frozen dates remain bound; otherwise unresolved"
            ),
        }
        result["reconciliation"] = reconcile(xlsx, all_records, date_skew_is_bound)
        observed = result["reconciliation"]["observed_delta_manifest"]
        delta_allowlist_matches = bool(
            full_run and _matches_delta_allowlist(
                observed["sha256"], observed["item_count"], observed["field_counts"])
        )
        result["delta_allowlist"] = {
            "discovery_mode": bool(args.discover_allowlist),
            "expected_sha256": EXPECTED_DELTA_MANIFEST_SHA256 or None,
            "expected_item_count": EXPECTED_DELTA_ITEM_COUNT,
            "expected_field_counts": EXPECTED_DELTA_FIELD_COUNTS,
            "matches": delta_allowlist_matches,
            "rule": (
                "only the exact field-specific delta set for the frozen XLSX and all "
                "twelve frozen PDF hashes is accepted; arbitrary additions/removals fail"
            ),
        }

        fatal_reasons = []
        if not result["xlsx_source"]["sha256_matches"]:
            fatal_reasons.append("authoritative XLSX hash changed")
        if len(xlsx["rows"]) != EXPECTED_XLSX_ROWS:
            fatal_reasons.append("authoritative XLSX row count changed")
        if result["pdf_sources"]["parse_residue_count"]:
            fatal_reasons.append("PDF parser left unclassified candidate-line residue")
        if result["reconciliation"]["unresolved_cells"]:
            fatal_reasons.append("printable cell differences are unclassified")
        if result["reconciliation"]["seg_order_inversion_count"]:
            fatal_reasons.append("PDF record order contradicts XLSX SEG_ORDER_ID")
        if result["reconciliation"]["unsafe_stream_attribution_count"]:
            fatal_reasons.append(
                "one or more source-delta characters cannot be assigned to one exact print field"
            )
        if not all(m["parameter_page_valid"] for m in metadata):
            fatal_reasons.append("one or more PDF parameter pages are invalid")
        if not all(m["sha256_matches"] for m in metadata):
            fatal_reasons.append("one or more authoritative PDF hashes changed")
        if not date_skew_is_bound:
            fatal_reasons.append("the frozen 09/08 XLSX to 09/15 PDF snapshot relation changed")
        if full_run:
            if len(metadata) != EXPECTED_PDF_MEMBERS:
                fatal_reasons.append("canonical PDF member count is not twelve")
            if total_pages != EXPECTED_PDF_PAGES:
                fatal_reasons.append("canonical PDF page count is not 4,123")
            if not args.discover_allowlist and not delta_allowlist_matches:
                fatal_reasons.append("observed real-delta manifest is not the exact allowlist")
        result["status"] = (
            "blocked" if fatal_reasons else
            "discovery_pass" if args.discover_allowlist else "pass"
        )
        result["fatal_reasons"] = fatal_reasons
        exit_code = 1 if fatal_reasons else 0
    except Exception as exc:
        result["status"] = "error"
        result["fatal_reasons"] = [f"{type(exc).__name__}: {exc}"]
        exit_code = 2
    finally:
        result["elapsed_seconds"] = round(time.monotonic() - started, 3)
        _write_json(args.result, result)
        print(f"result: {args.result}", flush=True)
        print(f"status: {result.get('status')} ({result['elapsed_seconds']:.1f}s)", flush=True)
        for reason in result.get("fatal_reasons", []):
            print(f"  - {reason}", flush=True)
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
