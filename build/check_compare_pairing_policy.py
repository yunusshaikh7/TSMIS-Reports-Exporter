"""Phase-3 E2 red/green gate for the approved duplicate-pairing policy.

This check intentionally does not import the Phase-3 independent oracle.  Its
small-matrix oracle below enumerates every rectangular assignment and applies
the approved smaller-side lexicographic tie rule directly.  That keeps this
gate independent of both ``compare_core`` and the statewide oracle.

The current (pre-E2) engine is expected to be RED here: it becomes greedy above
7!, flattens helper keys with ``|``, emits no exact/capped pairing trace, and
certifies above-cap results as complete.  Every production call is contained so
an exception becomes a named Checker failure rather than crashing the gate.

Run with the build venv:
    build/.venv/Scripts/python.exe build/check_compare_pairing_policy.py
"""
from __future__ import annotations

import argparse
import itertools
import random
import re
from collections.abc import Mapping

from _checklib import Checker, patch, scripts_path, temp_dir

scripts_path()

import compare_core as core
from openpyxl import load_workbook


PAIR_CAP = 100_000


def _validate_matrix(matrix):
    """Independent strict contract used by the exhaustive oracle."""
    if not isinstance(matrix, (list, tuple)) or not matrix:
        raise ValueError("matrix must contain at least one row")
    if not isinstance(matrix[0], (list, tuple)) or not matrix[0]:
        raise ValueError("matrix must contain at least one column")
    width = len(matrix[0])
    for row in matrix:
        if not isinstance(row, (list, tuple)) or len(row) != width:
            raise ValueError("matrix is ragged")
        for value in row:
            if type(value) is not int or value < 0:
                raise ValueError("costs must be non-negative integers")
    return len(matrix), width


def brute_assignment(matrix):
    """Return ``(total, smaller-side vector)`` by exhaustive enumeration.

    Side A defines the smaller side when dimensions are equal.  When B is
    smaller, the oriented oracle row ``b`` contains costs to every A index.
    Tuple comparison implements the approved lexicographically-smallest vector.
    """
    na, nb = _validate_matrix(matrix)
    if na <= nb:
        oriented = matrix
    else:
        oriented = tuple(tuple(matrix[a][b] for a in range(na))
                         for b in range(nb))
    small, large = len(oriented), len(oriented[0])
    best = None
    for vector in itertools.permutations(range(large), small):
        candidate = (sum(oriented[i][vector[i]] for i in range(small)), vector)
        if best is None or candidate < best:
            best = candidate
    return best


def _pairs_from_result(result):
    """Accept the legacy pair list and a future typed result's ``pairs`` seam."""
    pairs = getattr(result, "pairs", result)
    if not isinstance(pairs, (list, tuple)):
        raise TypeError(f"assignment result has no pair sequence: {type(result)!r}")
    answer = []
    for pair in pairs:
        if not isinstance(pair, (list, tuple)) or len(pair) != 2:
            raise ValueError(f"malformed assignment pair: {pair!r}")
        a, b = pair
        if type(a) is not int or type(b) is not int:
            raise ValueError(f"non-integer assignment pair: {pair!r}")
        answer.append((a, b))
    return answer


def _vector_from_pairs(na, nb, pairs):
    small_is_a = na <= nb
    small, large = (na, nb) if small_is_a else (nb, na)
    vector = [None] * small
    used_large = set()
    for a, b in pairs:
        if not (0 <= a < na and 0 <= b < nb):
            raise ValueError(f"out-of-range assignment pair {(a, b)!r}")
        i, j = (a, b) if small_is_a else (b, a)
        if vector[i] is not None or j in used_large:
            raise ValueError("assignment is not one-to-one")
        vector[i] = j
        used_large.add(j)
    if any(value is None for value in vector) or len(pairs) != small:
        raise ValueError("assignment does not cover the smaller side")
    return tuple(vector)


def production_solution(matrix):
    """Call production safely and return a normalized solution record."""
    try:
        na, nb = _validate_matrix(matrix)
        result = core._min_cost_pairs([list(row) for row in matrix])
        pairs = _pairs_from_result(result)
        vector = _vector_from_pairs(na, nb, pairs)
        total = sum(matrix[a][b] for a, b in pairs)
        return {"pairs": pairs, "vector": vector, "total": total, "error": None}
    except Exception as exc:  # the gate reports the defect; it must not crash
        return {"pairs": (), "vector": (), "total": None, "error": exc}


def legacy_greedy_score(matrix):
    """Reproduce only the documented old greedy scan to prove fixture bodies."""
    na, nb = _validate_matrix(matrix)
    flip = na > nb
    oriented = (matrix if not flip else
                tuple(tuple(matrix[a][b] for a in range(na)) for b in range(nb)))
    nr, nc = len(oriented), len(oriented[0])
    edges = sorted((oriented[r][col], r, col)
                   for r in range(nr) for col in range(nc))
    rows, cols, total = set(), set(), 0
    for value, row, col in edges:
        if row in rows or col in cols:
            continue
        rows.add(row)
        cols.add(col)
        total += value
        if len(rows) == nr:
            return total
    raise AssertionError("greedy fixture scan did not complete")


def hamming_matrix(side_a, side_b):
    if not side_a or not side_b:
        raise ValueError("Hamming fixture sides must be non-empty")
    width = len(side_a[0])
    if any(len(value) != width or set(value) - {"0", "1"}
           for value in tuple(side_a) + tuple(side_b)):
        raise ValueError("Hamming fixture rows must be equal-width binary strings")
    return tuple(tuple(sum(x != y for x, y in zip(a, b)) for b in side_b)
                 for a in side_a)


# The 4/6/4 body was recovered verbatim from the audit's preserved
# chunk12_pairing_nonmonotone_4ozja9ro/result.json and independently rebuilt
# here as Hamming distances.  The ledger records the same scores under
# CMP-AUD-008.
HAMMING_4_6_4_A = (
    "00000000000", "11100000000", "00000100000", "00000010000",
    "00000001000", "00000000100", "00000000010", "00000000001",
)
HAMMING_4_6_4_B = (
    "11000000000", "11111000000", "00000100000", "00000010000",
    "00000001000", "00000000100", "00000000010", "00000000001",
)

# The historical 14/10/8 body was not retained in the repository.  This exact
# equivalent was found by a deterministic xorshift32 search (seed 0x5eed1234,
# candidate 755011): B is a permutation of A plus eight distinct bit flips.
# Runtime exhaustive enumeration, not those historical scores, is the oracle.
HAMMING_14_10_8_A = (
    "111000110111", "011010001100", "100100100100", "000110110001",
    "000110100000", "011111010010", "011101101101", "011101111100",
)
HAMMING_14_10_8_B = (
    "111000010111", "011011001100", "100100101100", "000110100000",
    "000110110101", "011111011000", "011101111101", "011101100101",
)

# The historical 31/32 body was likewise unavailable.  This transparent
# equivalent embeds a 2x2 greedy trap [[4,0],[9,4]] beside six forced diagonal
# rows (costs 4,4,4,4,4,3); cross-block cost 10 keeps the exact diagonal total
# 31 while the greedy 0-edge forces total 32.
MATRIX_31_32_31 = (
    (4, 0, 10, 10, 10, 10, 10, 10),
    (9, 4, 10, 10, 10, 10, 10, 10),
    (10, 10, 4, 10, 10, 10, 10, 10),
    (10, 10, 10, 4, 10, 10, 10, 10),
    (10, 10, 10, 10, 4, 10, 10, 10),
    (10, 10, 10, 10, 10, 4, 10, 10),
    (10, 10, 10, 10, 10, 10, 4, 10),
    (10, 10, 10, 10, 10, 10, 10, 3),
)


def test_exact_against_bruteforce(c):
    print("\nD3 exact rectangular assignment against an exhaustive oracle:")
    rng = random.Random(0xD3E2)
    dimensions = ((1, 1), (1, 6), (2, 3), (3, 2), (3, 5),
                  (5, 3), (6, 7), (7, 6), (7, 7))
    mismatches = []
    for na, nb in dimensions:
        for sample in range(3):
            matrix = tuple(tuple(rng.randrange(8) for _ in range(nb))
                           for _ in range(na))
            expected = brute_assignment(matrix)
            actual = production_solution(matrix)
            if (actual["error"] is not None
                    or (actual["total"], actual["vector"]) != expected):
                mismatches.append((na, nb, sample, expected, actual))
    c.check("all deterministic <=7 rectangular matrices equal brute force",
            not mismatches, repr(mismatches[:3]))


def test_lexicographic_policy(c):
    print("\nD3 lexicographic smaller-side tie policy:")
    fixtures = (
        ("A-smaller 2x3", ((0, 0, 0), (0, 0, 0)), (0, 1)),
        ("B-smaller 3x2", ((0, 0), (0, 0), (0, 0)), (0, 1)),
        # Exactly two zero-cost perfect assignments: (1,2,0) and (2,0,1).
        # Choosing side B on this equal-size matrix would select the latter.
        ("equal 3x3 uses side A", ((9, 0, 0), (0, 9, 0), (0, 0, 9)),
         (1, 2, 0)),
    )
    for name, matrix, expected_vector in fixtures:
        oracle = brute_assignment(matrix)
        actual = production_solution(matrix)
        c.check(f"{name}: oracle fixture has the pinned lex vector",
                oracle[1] == expected_vector, f"oracle={oracle!r}")
        c.check(f"{name}: production chooses the pinned lex vector",
                actual["error"] is None and actual["vector"] == expected_vector,
                repr(actual))


def _paired_keys(result):
    if isinstance(result, (list, tuple)) and len(result) >= 2:
        return result[0], result[1]
    for a_name, b_name in (("keys_a", "keys_b"), ("keys_t", "keys_n")):
        if hasattr(result, a_name) and hasattr(result, b_name):
            return getattr(result, a_name), getattr(result, b_name)
    raise TypeError("pairing result has no paired key sequences")


def _hamming_rows(bits):
    return [["K", *value] for value in bits]


def test_retained_greedy_traps(c):
    print("\nRetained/equivalent 8x8 greedy traps:")
    fixtures = (
        ("14/10/8 deterministic Hamming equivalent",
         hamming_matrix(HAMMING_14_10_8_A, HAMMING_14_10_8_B), 14, 10, 8,
         HAMMING_14_10_8_A, HAMMING_14_10_8_B),
        ("4/6/4 recovered Hamming fixture",
         hamming_matrix(HAMMING_4_6_4_A, HAMMING_4_6_4_B), 4, 6, 4,
         HAMMING_4_6_4_A, HAMMING_4_6_4_B),
        ("31/32/31 constructed matrix", MATRIX_31_32_31, 31, 32, 31,
         None, None),
    )
    for name, matrix, positional, legacy, optimum, bits_a, bits_b in fixtures:
        oracle = brute_assignment(matrix)
        actual = production_solution(matrix)
        actual_positional = sum(matrix[i][i] for i in range(8))
        c.check(f"{name}: frozen body reproduces positional/greedy/optimal scores",
                (actual_positional, legacy_greedy_score(matrix), oracle[0])
                == (positional, legacy, optimum),
                repr((actual_positional, legacy_greedy_score(matrix), oracle)))
        c.check(f"{name}: production is exact with the oracle lex tie",
                actual["error"] is None
                and (actual["total"], actual["vector"]) == oracle,
                f"oracle={oracle!r}; production={actual!r}")
        c.check(f"{name}: production never exceeds positional cost",
                actual["error"] is None and actual["total"] <= positional,
                f"positional={positional}; production={actual!r}")

        # Both Hamming fixtures also traverse the real compared-cell cost and
        # occurrence-renumbering path; the direct matrix gate alone cannot prove
        # those seams are connected correctly.
        if bits_a is not None:
            header = ["Key"] + [f"Bit {i}" for i in range(len(bits_a[0]))]
            schema = core.CompareSchema(
                report_name="E2 Hamming", header=header,
                side_a="A", side_b="B", id_noun="row",
                id_noun_plural="rows", sides_noun="sides")
            rows_a, rows_b = _hamming_rows(bits_a), _hamming_rows(bits_b)
            keys_a = core.keys_for(rows_a, False)
            keys_b = core.keys_for(rows_b, False)
            try:
                paired_a, paired_b = _paired_keys(
                    core.pair_occurrences_by_similarity(
                        schema, rows_a, rows_b, keys_a, keys_b, False))
                counts = core.count_diffs(
                    schema, rows_a, rows_b, paired_a, paired_b,
                    core.union_keys(paired_a, paired_b), False)
                observed = counts["diff_cells"]
            except Exception as exc:
                observed = exc
            c.check(f"{name}: integrated row pairing has exact differing-cell cost",
                    observed == optimum,
                    f"expected={optimum}; observed={observed!r}")


def test_pairing_cost_semantics(c):
    print("\nPair cost consumes the approved asserting compared-cell state:")
    header = ["Key", "Ordinary", "Med-Wid", "Context", "Ditto", "BlankZero",
              "Literal marker"]
    schema = core.CompareSchema(
        report_name="E2 Costs", header=header, side_a="A", side_b="B",
        id_noun="row", id_noun_plural="rows", sides_noun="sides",
        medwid_fields=("Med-Wid",), context_fields=("Context",),
        ditto_nonasserting=True)
    base = ["K", "SAME", "6V", "same", "same", "same", "same"]
    mark = core._DIFF_MARK
    cases = (
        ("ordinary case remains significant", 1, {1: "ABC"}, {1: "abc"}),
        ("Med-Wid 06V/6V costs zero", 0, {2: "06V"}, {2: "6V"}),
        ("raw signed Med-Wid anomaly costs one", 1, {2: "-06V"}, {2: "-6V"}),
        ("context difference costs zero", 0, {3: "left"}, {3: "right"}),
        ("ditto difference costs zero", 0, {4: "+"}, {4: "other"}),
        ("blank and zero cost one", 1, {5: ""}, {5: 0}),
        ("equal literal marker content costs zero", 0,
         {6: f"A{mark}B"}, {6: f"A{mark}B"}),
        ("different literal marker content costs one", 1,
         {6: f"A{mark}B"}, {6: f"A{mark}C"}),
    )
    for name, expected, changes_a, changes_b in cases:
        row_a, row_b = list(base), list(base)
        for index, value in changes_a.items():
            row_a[index] = value
        for index, value in changes_b.items():
            row_b[index] = value
        try:
            actual = core._row_diff_count(schema, row_a, row_b, 0)
        except Exception as exc:
            actual = exc
        c.check(name, actual == expected,
                f"expected cost {expected}; production returned {actual!r}")


def test_malformed_matrix_rejection(c):
    print("\nMalformed assignment matrices fail closed:")
    fixtures = (
        ("empty matrix", []),
        ("zero-column matrix", [[]]),
        ("ragged matrix", [[0, 1], [1]]),
        ("negative cost", [[0, -1], [1, 0]]),
        ("Boolean cost", [[False, 1], [1, 0]]),
        ("fractional cost", [[0, 1.5], [1, 0]]),
    )
    for name, matrix in fixtures:
        try:
            core._min_cost_pairs(matrix)
        except (TypeError, ValueError):
            rejected = True
            detail = ""
        except Exception as exc:
            rejected = False
            detail = f"wrong exception type: {type(exc).__name__}: {exc}"
        else:
            rejected = False
            detail = "matrix was accepted"
        c.check(f"{name} is rejected with a contract error", rejected, detail)


def test_cap_boundary(c):
    print("\nProduct-cap boundary remains exact and genuinely rectangular:")
    # One row with 100,000 columns is the D3 square-padding guard.  Two zero-cost
    # columns also prove the lexicographically first minimum is selected.
    row = [1] * PAIR_CAP
    row[73] = row[-1] = 0
    try:
        result = core._min_cost_pairs([row])
        vector = _vector_from_pairs(1, PAIR_CAP, _pairs_from_result(result))
    except Exception as exc:
        vector = exc
    c.check("1x100000 assignment completes and chooses exact lex minimum",
            vector == (73,), f"observed={vector!r}")

    # Prove the duplicate-group product guard treats equality with the cap as
    # in-cap.  The spy keeps this entry-boundary check independent of the solver
    # check above while still exercising all 100,000 real compared-cell costs.
    schema = core.CompareSchema(
        report_name="E2 Cap", header=["Key", "Value"], side_a="A", side_b="B",
        id_noun="row", id_noun_plural="rows", sides_noun="sides")
    rows_a = [["K", "target"]]
    rows_b = [["K", "miss"]] * (PAIR_CAP - 1) + [["K", "target"]]
    keys_a = [("", "K", 1)]
    keys_b = [("", "K", i + 1) for i in range(PAIR_CAP)]
    seen = []

    def exact_boundary_spy(matrix, is_cancelled=None):
        seen.append((len(matrix), len(matrix[0])))
        if is_cancelled is not None and is_cancelled():
            raise AssertionError("boundary spy was unexpectedly cancelled")
        return [(0, PAIR_CAP - 1)]

    try:
        with patch(core, "_min_cost_pairs", exact_boundary_spy):
            paired_a, paired_b = _paired_keys(core.pair_occurrences_by_similarity(
                schema, rows_a, rows_b, keys_a, keys_b, False))
        paired_target = paired_b[-1][2] == paired_a[0][2]
    except Exception as exc:
        paired_target = exc
    c.check("1x100000 duplicate group enters the exact solver (not cap fallback)",
            seen == [(1, PAIR_CAP)] and paired_target is True,
            f"seen={seen!r}; paired_target={paired_target!r}")


def _get(value, *names, default=None):
    for name in names:
        if isinstance(value, Mapping) and name in value:
            return value[name]
        if hasattr(value, name):
            return getattr(value, name)
    return default


def _record_for_dimensions(records, na, nb):
    for record in records or ():
        if (_get(record, "side_a_size", "n_a", "na") == na
                and _get(record, "side_b_size", "n_b", "nb") == nb):
            return record
    return None


def _diag_is_complete(diag, na, nb, fallback_cost):
    if diag is None:
        return False
    return (
        _get(diag, "side_a_size", "n_a", "na") == na
        and _get(diag, "side_b_size", "n_b", "nb") == nb
        and _get(diag, "cap") == PAIR_CAP
        and _get(diag, "fallback", "fallback_policy") == "positional"
        and _get(diag, "fallback_cost", "observed_fallback_cost") == fallback_cost
    )


def test_public_pairing_results(c):
    print("\nPublic run/result carries exact traces and fail-closed cap state:")
    schema = core.CompareSchema(
        report_name="E2 Public", header=["Key", "Value"],
        side_a="A", side_b="B", id_noun="row", id_noun_plural="rows",
        sides_noun="sides")
    with temp_dir("tsmis_e2_pairing_") as tmp:
        # Within-cap all-zero 3x2 tie: side B is smaller and the pinned vector
        # is (0,1).  This probes persistence, not just the private solver.
        try:
            exact_result = core.run_compare(
                schema, [["K", "same"]] * 3, [["K", "same"]] * 2,
                False, tmp / "exact.xlsx", mode="values")
            exact_outcome = exact_result.comparison_outcome
            exact_trace = _record_for_dimensions(
                _get(exact_outcome, "pairing_trace", default=()), 3, 2)
        except Exception as exc:
            exact_result = exact_outcome = None
            exact_trace = exc
        c.check("public exact duplicate run reports pairing_quality=exact",
                exact_outcome is not None
                and _get(exact_outcome, "pairing_quality") == "exact",
                repr(_get(exact_outcome, "pairing_quality")))
        c.check("public exact duplicate run persists the pinned B-side trace",
                exact_trace is not None and not isinstance(exact_trace, Exception)
                and _get(exact_trace, "smaller_side") in ("b", "B")
                and tuple(_get(exact_trace, "assignment_vector", default=())) == (0, 1)
                and _get(exact_trace, "total_cost") == 0
                and _get(exact_trace, "quality") == "exact"
                and _get(exact_trace, "exact") is True,
                repr(exact_trace))

        # 317*316 = 100,172, the approved above-cap regression.  Positional
        # fallback yields 316 diffs and one A-only row; those useful counts are
        # retained, but they must be explicitly non-certifying.
        rows_a = [["K", f"V{i}"] for i in range(317)]
        rows_b = [["K", f"V{i}"] for i in reversed(range(316))]
        try:
            capped_result = core.run_compare(
                schema, rows_a, rows_b, False, tmp / "317x316.xlsx", mode="values")
            capped_outcome = capped_result.comparison_outcome
            capped_diag = _record_for_dimensions(
                _get(capped_outcome, "capped_group_diagnostics", default=()),
                317, 316)
            capped_counts = _get(capped_outcome, "counts")
            capped_book = load_workbook(
                tmp / "317x316.xlsx", data_only=False, read_only=True)
            capped_banner = capped_book["Summary"]["B3"].value
            capped_book.close()
        except Exception as exc:
            capped_result = capped_outcome = capped_counts = None
            capped_diag = exc
            capped_banner = exc
        c.check("317x316 fallback counts remain observable",
                capped_counts is not None
                and _get(capped_counts, "paired_rows") == 316
                and _get(capped_counts, "side_a_only_rows") == 1
                and _get(capped_counts, "side_b_only_rows") == 0
                and _get(capped_counts, "differing_cells") == 316,
                repr(capped_counts))
        c.check("317x316 public result is partial with pairing_quality=capped",
                capped_result is not None
                and _get(capped_result, "completion") == "partial"
                and _get(capped_outcome, "completion") == "partial"
                and _get(capped_outcome, "pairing_quality") == "capped",
                (f"legacy completion={_get(capped_result, 'completion')!r}; "
                 f"typed completion={_get(capped_outcome, 'completion')!r}; "
                 f"quality={_get(capped_outcome, 'pairing_quality')!r}"))
        c.check("317x316 emits the structured positional fallback diagnostic",
                _diag_is_complete(capped_diag, 317, 316, 316), repr(capped_diag))
        c.check("317x316 never presents fallback artifacts as certified differences",
                isinstance(capped_banner, str)
                and capped_banner.startswith("=IF(")
                and "REGENERATE REQUIRED" in capped_banner
                and "✗ PARTIAL / PAIRING LIMIT" in capped_banner
                and "diagnostic counts, not certified differences" in capped_banner
                and "DIFFERENCES FOUND" not in capped_banner
                and capped_result.summary_lines[0].startswith(
                    "⚠ PARTIAL / PAIRING LIMIT")
                and "not certified differences" in capped_result.summary_lines[0]
                and "DIFFERENCES FOUND" not in capped_result.summary_lines[0],
                f"banner={capped_banner!r}; lines={capped_result.summary_lines!r}")

        # Even a clean-looking positional result above the cap cannot certify a
        # match: identity remains unproved.  This is the direct no-green case.
        same = [["K", "same"]] * 317
        try:
            no_green_result = core.run_compare(
                schema, same, same, False, tmp / "317x317.xlsx", mode="values")
            no_green_outcome = no_green_result.comparison_outcome
            no_green_diag = _record_for_dimensions(
                _get(no_green_outcome, "capped_group_diagnostics", default=()),
                317, 317)
            no_green_book = load_workbook(
                tmp / "317x317.xlsx", data_only=False, read_only=True)
            no_green_banner = no_green_book["Summary"]["B3"].value
            no_green_book.close()
        except Exception as exc:
            no_green_result = no_green_outcome = None
            no_green_diag = exc
            no_green_banner = exc
        c.check("317x317 clean fallback is partial/capped and never match/green",
                no_green_result is not None
                and _get(no_green_result, "completion") == "partial"
                and _get(no_green_outcome, "completion") == "partial"
                and _get(no_green_outcome, "pairing_quality") == "capped"
                and _get(no_green_result, "verdict") != "match"
                and _get(no_green_outcome, "verdict") != "match",
                (f"legacy=({_get(no_green_result, 'completion')!r}, "
                 f"{_get(no_green_result, 'verdict')!r}); "
                 f"typed=({_get(no_green_outcome, 'completion')!r}, "
                 f"{_get(no_green_outcome, 'verdict')!r}, "
                 f"{_get(no_green_outcome, 'pairing_quality')!r})"))
        c.check("317x317 emits a zero-cost structured capped diagnostic",
                _diag_is_complete(no_green_diag, 317, 317, 0),
                repr(no_green_diag))
        c.check("317x317 workbook Summary is visibly non-certifying",
                isinstance(no_green_banner, str)
                and no_green_banner.startswith("=IF(")
                and "REGENERATE REQUIRED" in no_green_banner
                and "✗ PARTIAL / PAIRING LIMIT" in no_green_banner
                and "EVERYTHING MATCHES" not in no_green_banner,
                repr(no_green_banner))
        c.check("317x317 returned summary is visibly non-certifying",
                no_green_result is not None
                and no_green_result.summary_lines
                and no_green_result.summary_lines[0].startswith(
                    "⚠ PARTIAL / PAIRING LIMIT")
                and all("EVERYTHING MATCHES" not in line
                        for line in no_green_result.summary_lines),
                repr(_get(no_green_result, "summary_lines")))


def _excel_rebuild(path):
    import win32com.client as win32
    excel = book = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        book = excel.Workbooks.Open(
            str(path.resolve()), UpdateLinks=0, ReadOnly=False)
        excel.CalculateFullRebuild()
        book.Save()
        book.Close(SaveChanges=False)
        book = None
    finally:
        if book is not None:
            book.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()


def test_helper_key_injection(c, run_excel=False):
    print("\nWorkbook helper identities are injective when components contain pipes:")
    schema = core.CompareSchema(
        report_name="E2 Pipes", header=["Key", "Value"], side_a="A", side_b="B",
        id_noun="row", id_noun_plural="rows", sides_noun="sides")
    side_a = [["R|X", "K", "A1"], ["R", "X|K", "A2"],
              ["R|ONLY", "K|ONLY", "A3"]]
    side_b = [["R|X", "K", "A1"], ["R", "X|K", "B2"]]
    with temp_dir("tsmis_e2_pipe_") as tmp:
        output = tmp / "pipe.xlsx"
        workbook = None
        try:
            result = core.run_compare(
                schema, side_a, side_b, True, output, mode="formulas")
            # This fixture is tiny.  Load it eagerly so no streaming worksheet
            # iterator can retain a Windows read handle while installed Excel
            # later replaces the workbook during Save().
            workbook = load_workbook(output, data_only=False, read_only=False)
            sheet = workbook["A"]
            rows = sheet.iter_rows(values_only=True)
            header = next(rows)
            helper_index = header.index("Key (helper)")
            helpers = [next(rows)[helper_index] for _ in range(3)]
            comparison = workbook["Comparison"]
            comparison_rows = comparison.iter_rows(values_only=True)
            comparison_header = next(comparison_rows)
            a_row_index = comparison_header.index("A Row")
            b_row_index = comparison_header.index("B Row")
            lookup_formulas = []
            for comparison_row in comparison_rows:
                lookup_formulas.extend(
                    formula for formula in (
                        comparison_row[a_row_index], comparison_row[b_row_index])
                    if isinstance(formula, str) and formula.startswith("="))
            only_a = workbook["Only in A"]
            only_rows = only_a.iter_rows(values_only=True)
            only_header = next(only_rows)
            only_formula = next(only_rows)[only_header.index("A Row")]
            counts = result.comparison_outcome.counts
        except Exception as exc:
            helpers = exc
            lookup_formulas = ()
            only_formula = None
            counts = None
        finally:
            if workbook is not None:
                workbook.close()
        c.check("pipe-bearing tuple identities remain distinct in workbook helpers",
                isinstance(helpers, list) and len(set(helpers)) == 3
                and all(str(value).startswith("__CMP_E2_KEY_V1_")
                        for value in helpers),
                f"helpers={helpers!r}")
        tokens = [
            match.group(1)
            for formula in lookup_formulas
            for match in re.finditer(
                r'MATCH\("(__CMP_E2_KEY_V1_\d+)"', formula)
        ]
        c.check("Comparison MATCH formulas use distinct opaque literals only",
                len(set(tokens)) == 3
                and all('&"|"&' not in formula for formula in lookup_formulas),
                f"tokens={tokens!r}; formulas={lookup_formulas!r}")
        only_tokens = (re.findall(
            r'MATCH\("(__CMP_E2_KEY_V1_\d+)"', only_formula)
            if isinstance(only_formula, str) else [])
        c.check("Only-in MATCH formulas use the same opaque identity seam",
                len(set(only_tokens)) == 1
                and only_tokens[0] in set(helpers)
                and '&"|"&' not in only_formula,
                f"tokens={only_tokens!r}; formula={only_formula!r}")
        c.check("pipe fixture's typed Python result still exposes its real diff",
                counts is not None and _get(counts, "differing_cells") == 1
                and _get(counts, "side_a_only_rows") == 1,
                repr(counts))
        if run_excel:
            calculated = None
            try:
                _excel_rebuild(output)
                calculated = load_workbook(
                    output, data_only=True, read_only=True)
                comparison_rows = calculated["Comparison"].iter_rows(
                    values_only=True)
                calculated_header = next(comparison_rows)
                status_col = calculated_header.index("Status")
                diffs_col = calculated_header.index("Diffs")
                calculated_rows = list(comparison_rows)
                statuses = [row[status_col] for row in calculated_rows]
                diffs = [row[diffs_col] for row in calculated_rows]
                summary = calculated["Summary"]
                checks = [row[2] for row in summary.iter_rows(values_only=True)
                          if len(row) > 2 and isinstance(row[1], str)
                          and isinstance(row[2], str)
                          and row[2] in ("OK", "CHECK")]
                excel_error = None
            except Exception as exc:
                statuses = diffs = checks = ()
                excel_error = f"{type(exc).__name__}: {exc}"
            finally:
                if calculated is not None:
                    calculated.close()
            c.check("installed Excel rebuild preserves all pipe-bearing identities",
                    excel_error is None
                    and statuses == ["Both", "Both", "A only"]
                    and diffs == [0, 1, None],
                    f"error={excel_error!r}; statuses={statuses!r}; diffs={diffs!r}")
            c.check("installed Excel pipe fixture passes every Summary self-check",
                    len(checks) >= 6 and set(checks) == {"OK"},
                    f"error={excel_error!r}; checks={checks!r}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--excel", action="store_true",
        help="run installed Excel full-rebuild on the opaque pipe fixture")
    args = parser.parse_args()
    c = Checker()
    print("Phase-3 E2 duplicate-pairing policy gate")
    print("Oracle: independent exhaustive rectangular enumeration (small fixtures)")
    test_exact_against_bruteforce(c)
    test_lexicographic_policy(c)
    test_retained_greedy_traps(c)
    test_pairing_cost_semantics(c)
    test_malformed_matrix_rejection(c)
    test_cap_boundary(c)
    test_public_pairing_results(c)
    test_helper_key_injection(c, run_excel=args.excel)
    return c.summary()


if __name__ == "__main__":
    raise SystemExit(main())
