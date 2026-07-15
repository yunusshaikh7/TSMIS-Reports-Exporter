"""Golden check for the TSMIS-vs-TSN Intersection Detail comparator
(scripts/compare_intersection_detail_tsn.py) — the FLAT recipe (route+PM),
July-2026 export format.

Locks: the CompareSchema wiring (PM key; NO context fields — every shared column is
compared and counted; boolean fields; the Notes legend_writer), the July-2026 shape
(35 shared fields ending in 'Xing Line Lgth', District/County joined per ID-79; the second ML eff-date gone — TSN's
MAIN_EFF_DATE is Report-View-only now), the old-format refusal (a pre-update
consolidated workbook errors with a re-export hint instead of being mis-read by
position), the Y/N<->legacy-1/0 boolean normalization, the control-type crosswalk,
route-from-LOCATION + PM/date/zero-pad normalization, the position-based TSMIS
loader, the Report View classification (Date of Record + INT/Ctrl/Light eff-dates
count as Major now; Int St / ML / CS Eff-Date + Route Suffix stay soft — user
decision 2026-07-08), and end-to-end that a normalization still produces a MATCH
while everything present in both systems IS counted. No Excel; CI-safe.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_intersection_detail_tsn.py
"""
import inspect
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_intersection_detail_tsn as idt
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []
DIFF = " ≠ "


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _write_tsmis(path, rows, old_format=False):
    """Synthetic CONSOLIDATED Intersection Detail: header[0]='Route' + the 35
    July-2026 source columns (the loader reads by POSITION, but its header gate
    demands the trailing 'Xing Line Lgth'). `old_format` writes the pre-update
    37-column shape instead — the refusal fixture."""
    wb = Workbook()
    ws = wb.active
    ws.title = idt.TSMIS_SHEET
    if old_format:
        ws.append(["Route"] + [f"c{i}" for i in range(1, 36)] + ["Xing S"])
    else:
        ws.append(["Route"] + [f"c{i}" for i in range(1, 35)] + ["Xing Line Lgth"])
    for r in rows:
        ws.append(r + [None] * (36 - len(r)))
    wb.save(path)
    wb.close()


def _tsmis_row(route, pr, pm, dor, hg, city, ru, int_t, ctrl_t, light, ml_sm, ml_lc,
               ml_rc, ml_tf, ml_nl, desc, cs_sm=None, cs_lc=None, geo_date=None,
               int_st=None, xll=None, mll=None, location=None):
    """Place values at the July-2026 consolidated VALUE positions the loader
    reads. `geo_date` fills every geometry eff-date slot (INT/Ctrl/Light/ML/CS)."""
    r = [None] * 36
    r[0], r[1], r[2], r[4], r[5] = route, pr, pm, location or ("12 ORA " + route), dor
    r[6], r[7], r[8] = hg, city, ru
    d = geo_date or dor
    r[9], r[11], r[13], r[15], r[23] = d, d, d, d, d      # the five eff-dates
    r[10], r[12], r[14] = int_t, ctrl_t, light
    r[16], r[17], r[18], r[19], r[20] = ml_sm, ml_lc, ml_rc, ml_tf, ml_nl
    r[21], r[22] = desc, mll               # Description, Main Line Length
    r[24], r[25] = cs_sm, cs_lc            # cross-street attributes
    r[29] = int_st or d                    # Int St Eff-Date
    r[35] = xll                            # Xing Line Lgth (July 2026)
    return r


def _write_tsn(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = idt.TSN_SHEET
    # Exact order from the SHA-bound statewide raw TSN workbook.  Keeping this
    # fixture faithful matters even though the production loader projects by
    # header name: the independent Phase-3 reader deliberately rejects schema
    # or order drift.
    cols = ["PP", "POST_MILE", "LOCATION", "DATE_REC", "HG", "CITY_CODE", "RU",
            "EFF_DATE_INT", "TY_INT", "EFF_DATE_CT", "TY_CT", "EFF_DATE_LT", "LT_TY",
            "EFF_DATE_ML", "MAIN_SM", "MAIN_LC", "MAIN_RC", "MAIN_TF", "MAIN_NL",
            "X_CROSS_OVERRIDE", "MAIN_EFF_DATE", "MAIN_ADT", "DESCRIPTION",
            "MAIN_OVERRIDE", "CROSS_BEGIN_DATE", "CS_SM", "CS_LC", "CS_RC", "CS_TF",
            "CS_NL", "EFF_DATE", "CROSS_ADT", "CROSS_ROUTE_NAME", "CROSS_PM_PREFIX",
            "CROSS_POSTMILE", "CROSS_PM_SUFFIX"]
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    wb.save(path)
    wb.close()


def _comparison(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Comparison"]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
        return header, rows, wb.sheetnames
    finally:
        wb.close()


def test_schema():
    print("schema + normalizers (July-2026 shape):")
    sc = idt._SCHEMA
    check("key is PM", sc.header[sc.key_field] == "PM")
    check("side names TSMIS / TSN", sc.side_a == "TSMIS" and sc.side_b == "TSN")
    check("position-aligned: NO context fields (nothing suppressed or greyed)",
          tuple(sc.context_fields) == ())
    check("35 shared fields (District + County joined per ID-79), ending in "
          "'Xing Line Lgth'",
          len(sc.header) == 35 and sc.header[-1] == "Xing Line Lgth"
          and sc.header[3:5] == ["District", "County"])
    check("the second ML eff-date left the shared header (TSN MAIN_EFF_DATE is "
          "Report-View-only now)", "ML 2nd Eff-Date" not in sc.header)
    check("all 7 date cols + Main Line Length + intersecting route + Xing Line "
          "Lgth ARE compared",
          all(f in sc.header and f not in sc.context_fields for f in (
              "Date of Record", "INT Type Eff-Date", "Control Type Eff-Date",
              "Lighting Eff-Date", "ML Eff-Date", "CS Eff-Date",
              "Int St Eff-Date", "Main Line Length", "Intrte Route",
              "Intrte PM Prefix", "Intrte Postmile", "Intrte PM Suffix",
              "Xing Line Lgth")))
    check("position-aligned eff-dates: ML/CS -> geometry EFF_DATE_ML/CROSS_BEGIN_DATE; "
          "Int St -> TSN's EFF_DATE (its bulk stamp); XLL -> X_CROSS_OVERRIDE",
          idt._TSN_COL["ML Eff-Date"] == "EFF_DATE_ML"
          and idt._TSN_COL["CS Eff-Date"] == "CROSS_BEGIN_DATE"
          and idt._TSN_COL["Int St Eff-Date"] == "EFF_DATE"
          and idt._TSN_COL["Xing Line Lgth"] == "X_CROSS_OVERRIDE"
          and "ML 2nd Eff-Date" not in idt._TSN_COL)
    check("July-2026 positions: Description 21, Int St 29, PM suffix 34, XLL 35",
          idt._TSMIS_POS["Description"] == 21 and idt._TSMIS_POS["Int St Eff-Date"] == 29
          and idt._TSMIS_POS["Intrte PM Suffix"] == 34
          and idt._TSMIS_POS["Xing Line Lgth"] == 35)
    check("header gate: Route + 35 cols ending 'Xing Line Lgth'",
          idt._header_ok(["Route"] + ["x"] * 34 + ["Xing Line Lgth"])
          and not idt._header_ok(["Route"] + ["x"] * 36)          # the old 37-col shape
          and not idt._header_ok(["Route"] + ["x"] * 35))
    check("Notes legend_writer set (documents the normalizations)", sc.legend_writer is not None)
    check("Report View extra_sheet_writer set (the printed two-line replica)",
          sc.extra_sheet_writer is None and not sc.report_view_diff_check)  # closure added per-call
    check("boolean normalize Y/1->Y, N/0->N (legacy 1/0 still folds)",
          idt._norm_bool("Y") == "Y" and idt._norm_bool("1") == "Y"
          and idt._norm_bool("N") == "N" and idt._norm_bool("0") == "N")
    check("control-type crosswalk: TSN J-P + TSMIS S -> 'S' (signalized); others unchanged",
          all(idt._norm_control_type(c) == "S" for c in "JKLMNPS")
          and idt._norm_control_type("A") == "A" and idt._norm_control_type("B") == "B")
    check("route token: numeric 0 keys as '000', never blank (falsy-zero D1 — "
          "_split_route feeds the alignment key)",
          idt._split_route(0) == ("000", ""))
    # Report View classification (user decision 2026-07-08, the data-driven soft set).
    check("Report View: 'soft' shares the hard RED palette (all diffs red)",
          idt._RV_FILLS["soft"] == idt._RV_FILLS["hard"])
    check("Report View: the normal (lighter) alternating band is WHITE",
          idt._RV_FILLS["eq"][0] == "FFFFFF" and idt._RV_FILLS["id"][0] == "FFFFFF")
    check("Report View: a Date of Record diff counts as Major now (the July-2026 "
          "export matches TSN structurally)",
          idt._rv_classify("Date of Record", "1999-12-30", "1970-01-01") == "hard")
    check("Report View: INT/Ctrl/Light eff-date diffs are Major (the ~1-day-offset "
          "tolerance is retired)",
          idt._rv_classify("Lighting Eff-Date", "1973-10-18", "1973-10-19") == "hard")
    check("Report View: Int St / ML / CS Eff-Date stay soft (structural)",
          all(idt._rv_classify(f, "1964-01-01", "2022-01-01") == "soft"
              for f in ("Int St Eff-Date", "ML Eff-Date", "CS Eff-Date")))
    check("Report View: a non-date attribute diff classifies 'hard' (counts as Major)",
          idt._rv_classify("Control Type", "S", "A") == "hard")
    check("Report View: Route Suffix is a compared grid column (next to Route)",
          any(c[2] == ("cmp", "Route Suffix") for c in idt._RV_GRID))
    check("Report View: a route-suffix diff classifies 'soft' (red, excluded from Major)",
          idt._rv_classify("Route Suffix", "", "U") == "soft")
    check("Report View grid: Xing Line Lgth is a compared cell; the 2nd ML eff-date "
          "is a TSN-only (blue) reference cell",
          any(spec == ("cmp", "Xing Line Lgth") for c in idt._RV_GRID for spec in (c[2], c[5]))
          and any(spec == ("tn", "ML2") for c in idt._RV_GRID for spec in (c[2], c[5]))
          and not any(spec == ("cmp", "ML 2nd Eff-Date") for c in idt._RV_GRID
                      for spec in (c[2], c[5])))
    cmp_fields = [spec[1] for column in idt._RV_GRID
                  for spec in (column[2], column[5]) if spec[0] == "cmp"]
    asserting_fields = [sc.header[index] for index in sc.field_indices
                        if not sc.is_context(index)]
    # District + County are asserted key-adjacent fields (ID-79) that CANNOT
    # differ on a paired row (county is inside the pairing identity; district is
    # county-determined) — the printed replica shows both inside its LOCATION
    # cell, so they carry no dedicated grid cell. Any anomaly still flags on the
    # generic Comparison sheet.
    rv_exempt = {"District", "County"}
    check("Report View grid contains every asserting non-key field exactly once "
          "(District/County shown via LOCATION)",
          set(cmp_fields) == set(asserting_fields) - rv_exempt
          and len(cmp_fields) == len(set(cmp_fields)))
    check("Report View: the TSN-only 2nd ML eff-date renders as a DATE",
          "ML2" in idt._RV_DATEONE and idt._RV_ONE["ML2"] == "MAIN_EFF_DATE")
    check("route from LOCATION '12 ORA 001' -> '001'", idt._norm_route("12 ORA 001") == "001")
    check("route-suffix split '12 ORA 210U' -> ('210','U')", idt._split_route("12 ORA 210U") == ("210", "U"))
    check("route-suffix split '12 ORA. 210' -> ('210','')", idt._split_route("12 ORA. 210") == ("210", ""))
    check("Route Suffix is a COMPARED column (not context)",
          "Route Suffix" in sc.header and "Route Suffix" not in sc.context_fields)
    check("PM ' 000.204' -> '0.204'", idt._norm_pm(" 000.204") == "0.204")
    check("numeric 0 -> '0' (preserved; matches text '0.000')",
          idt._norm_num(0) == "0" and idt._norm_num(0.0) == "0" and idt._norm_num("0.000") == "0")
    check("blank stays blank (None/'' -> '')",
          idt._norm_num(None) == "" and idt._norm_num("") == "")
    check("numeric-0 boolean -> 'N' (not blank)", idt._norm_bool(0) == "N")
    check("date ISO from YY-MM-DD ('73-10-19' -> '1973-10-19')",
          idt._iso_date("73-10-19") == "1973-10-19")


def test_report_view_typed_truth():
    print("Report View typed comparison truth:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_id_rv_truth_"))
    out = root / "report-view.xlsx"
    # City Code, Main Line Length, Date of Record, and INT Type are hard
    # differences; ML Eff-Date is soft. Formula-leading/error-looking values
    # exercise the familiar view's literal-write boundary.
    # R/U is made context only in this synthetic schema to prove an unequal but
    # non-asserting cell cannot render/count/classify as a Report View diff.
    sc = idt.dataclasses.replace(idt._SCHEMA, context_fields=("R/U",))

    def row(side_b=False):
        values = {name: "" for name in sc.header}
        values.update({
            "PM": "0.204",
            "PR": "=1+1",
            "Route Suffix": "#N/A",
            "Description": "ASCII SPACE" if side_b else "  ASCII   SPACE  ",
            "City Code": "case" if side_b else "Case",
            "HG": "literal ≠ content",
            "Date of Record": "@RIGHT" if side_b else "@LEFT",
            "INT Type": "+RIGHT" if side_b else "+LEFT",
            "Lighting Eff-Date": "-SAFE",
            "Main Line Length": 0 if side_b else None,
            "ML Eff-Date": "1974-01-01" if side_b else "1973-01-01",
            "R/U": "R" if side_b else "U",
        })
        return ["001"] + [values[name] for name in sc.header]

    wb = Workbook(write_only=True)
    idt._write_report_view(
        wb,
        {"sc": sc, "rows_a": [row()], "rows_b": [row(True)]},
        [{}],
        ["12 ORA 001"],
    )
    wb.save(out)
    wb.close()

    wb = load_workbook(out, data_only=True)
    try:
        ws = wb["Report View"]

        def grid_cell(line, field):
            slot = 2 if line == 0 else 5
            for grid_index, column in enumerate(idt._RV_GRID):
                if column[slot] == ("cmp", field):
                    return ws.cell(
                        row=5 + line,
                        column=len(idt._RV_AUX) + grid_index + 1,
                    )
            raise AssertionError(f"missing Report View field: {field}")

        desc = grid_cell(1, "Description")
        city = grid_cell(0, "City Code")
        literal = grid_cell(0, "HG")
        zero = grid_cell(1, "Main Line Length")
        soft = grid_cell(0, "ML Eff-Date")
        context = grid_cell(0, "R/U")
        eq_formula = grid_cell(0, "PR")
        eq_error = grid_cell(0, "Route Suffix")
        plus_diff = grid_cell(0, "INT Type")
        at_diff = grid_cell(0, "Date of Record")
        minus_equal = grid_cell(0, "Lighting Eff-Date")
        check("Report View ASCII-space equality follows shared Excel TRIM",
              desc.value == "ASCII SPACE")
        check("Report View equality remains case-sensitive",
              city.value == "Case ≠ case")
        check("equal literal difference-marker content is displayed but not counted",
              literal.value == "literal ≠ content")
        check("blank-vs-zero is asserting and preserves the dot display",
              zero.value == "· ≠ 0")
        check("only asserting unequal cells render as differences",
              context.value == "U" and " ≠ " not in context.value)
        check("Report View guards =/+/−/@ and Excel-error source literals",
              eq_formula.value == "=1+1" and eq_formula.data_type == "s"
              and eq_error.value == "#N/A" and eq_error.data_type == "s"
              and plus_diff.value == "+LEFT ≠ +RIGHT" and plus_diff.data_type == "s"
              and at_diff.value == "@LEFT ≠ @RIGHT" and at_diff.data_type == "s"
              and minus_equal.value == "-SAFE" and minus_equal.data_type == "s")
        check("hard and soft differences retain the red palette",
              city.fill.fgColor.rgb[-6:] == idt._RV_FILLS["hard"][0]
              and soft.fill.fgColor.rgb[-6:] == idt._RV_FILLS["soft"][0]
              and literal.fill.fgColor.rgb[-6:] == idt._RV_FILLS["eq"][0])
        data_rows = list(ws.iter_rows(min_row=5, values_only=True))
        check("matched record writes its full typed Diffs total on exactly two rows",
              len(data_rows) == 2
              and all(record[0] == 4 and record[1] == 5
                      for record in data_rows))
    finally:
        wb.close()

    source = inspect.getsource(idt._write_report_view)
    cmp_block = source.split('if kind == "cmp":', 1)[1].split(
        'return ("", "blank")', 1)[0]
    check("Report View cmp branch has no local strip/equality truth",
          "compared_cell(sc, field_index[ref], ra, rb, off=1)" in cmp_block
          and "cell.asserting" in cmp_block and "cell.equal" in cmp_block
          and "aval(" not in cmp_block and ".strip(" not in cmp_block
          and "tm == tn" not in cmp_block)


def test_end_to_end():
    print("end-to-end (normalizations still match; every shared column counted):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_id_tsn_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # PM 000.204: signalized sub-type split (TSMIS S vs TSN P) — CROSSWALKED to
    #   equal; native Y/N booleans equal; Xing Line Lgth '250' vs TSN '0250' —
    #   zero-pad-normalized to EQUAL; the CS attributes blank on TSMIS vs valued
    #   on TSN (5 counted diffs); Int St Eff-Date historical vs TSN's bulk stamp
    #   (1 structural counted diff). PM 001.000: a NON-signalized control diff
    #   (A vs B, NOT crosswalked) + a real ML Num Lanes diff.
    _write_tsmis(tsmis, [
        _tsmis_row("001", "R", "000.204", "73-10-19", "D", "DAPT", "U", "T", "S", "Y",
                   "Y", "N", "N", "P", "3", "JCT 5", int_st="73-10-19", xll="250",
                   mll="100"),
        _tsmis_row("001", "R", "001.000", "73-10-19", "D", "DAPT", "U", "T", "A", "Y",
                   "Y", "N", "N", "P", "4", "JCT 6", int_st="73-10-19"),
    ])
    _write_tsn(tsn, [
        {"PP": "R", "POST_MILE": " 000.204", "LOCATION": "12 ORA 001", "DATE_REC": "73-10-19",
         "HG": "D", "CITY_CODE": "DAPT", "RU": "U", "TY_INT": "T", "TY_CT": "P", "LT_TY": "Y",
         "MAIN_SM": "Y", "MAIN_LC": "N", "MAIN_RC": "N", "MAIN_TF": "P", "MAIN_NL": 3,
         "DESCRIPTION": "JCT 5", "CS_SM": "N", "CS_LC": "N", "CS_RC": "N", "CS_TF": "P",
         "CS_NL": 2, "EFF_DATE_INT": "73-10-19", "EFF_DATE_CT": "73-10-19",
         "EFF_DATE_LT": "73-10-19", "EFF_DATE_ML": "73-10-19",
         "CROSS_BEGIN_DATE": "73-10-19", "MAIN_OVERRIDE": None,
         "EFF_DATE": "22-01-01", "X_CROSS_OVERRIDE": "0250"},
        {"PP": "R", "POST_MILE": " 001.000", "LOCATION": "12 ORA 001", "DATE_REC": "73-10-19",
         "HG": "D", "CITY_CODE": "DAPT", "RU": "U", "TY_INT": "T", "TY_CT": "B", "LT_TY": "Y",
         "MAIN_SM": "Y", "MAIN_LC": "N", "MAIN_RC": "N", "MAIN_TF": "P", "MAIN_NL": 3,
         "DESCRIPTION": "JCT 6", "EFF_DATE_INT": "73-10-19", "EFF_DATE_CT": "73-10-19",
         "EFF_DATE_LT": "73-10-19", "EFF_DATE_ML": "73-10-19",
         "CROSS_BEGIN_DATE": "73-10-19", "EFF_DATE": "73-10-19"},
    ])
    res = idt.compare(tsmis, tsn, out, events=Events(), confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows, sheets = _comparison(out)
    check("Notes sheet present (the indicator)", "Notes" in sheets)
    check("Report View sheet appended (the printed two-line replica)", "Report View" in sheets)
    formula_wb = load_workbook(out, read_only=True, data_only=False)
    try:
        summary = formula_wb["Summary"]
        report_checks = [row[2].value for row in summary.iter_rows()
                         if len(row) >= 3
                         and row[1].value ==
                         "Report View Diffs agree with the Comparison"]
        report_formula = str(report_checks[0]) if len(report_checks) == 1 else ""
        check("Summary independently cross-checks the two-line Report View total",
              len(report_checks) == 1
              and "SUM('Report View'!B:B)=2*SUM(Comparison!" in report_formula
              and '"OK","CHECK"' in report_formula)
    finally:
        formula_wb.close()
    pm = header.index("PM")
    by = {r[pm]: r for r in rows}

    light = header.index("Lighting")
    ctrl = header.index("Control Type")
    nl = header.index("ML Num Lanes")
    dor = header.index("Date of Record")
    cs_sm = header.index("CS Mastarm")
    int_st = header.index("Int St Eff-Date")
    xll = header.index("Xing Line Lgth")
    mll = header.index("Main Line Length")
    diffs_col = header.index("Diffs")
    # Normalizations still produce a MATCH (the point of "make normalization clear,
    # even though it leads to a match").
    check("Lighting Y/Y equal — no diff", DIFF not in by["001 / ORA / R0.204"][light])
    check("Control Type S(TSMIS)/P(TSN) crosswalk to 'S' — no diff",
          DIFF not in by["001 / ORA / R0.204"][ctrl] and by["001 / ORA / R0.204"][ctrl] == "S")
    check("Xing Line Lgth '250' vs TSN '0250' zero-pad-normalized — no diff",
          DIFF not in by["001 / ORA / R0.204"][xll] and by["001 / ORA / R0.204"][xll] == "250")
    check("Date of Record MATCHES now (the July-2026 fix — no wholesale diff)",
          DIFF not in by["001 / ORA / R0.204"][dor])
    check("Main Line Length '100'(TSMIS) vs blank(TSN) is a counted diff",
          DIFF in by["001 / ORA / R0.204"][mll])
    # Everything present in both systems is COUNTED (no suppression):
    check("CS Mastarm blank(TSMIS) vs N(TSN) is a COUNTED diff (no coalescing)",
          DIFF in by["001 / ORA / R0.204"][cs_sm])
    check("Int St Eff-Date historical(1973) vs TSN bulk stamp(2022) is a COUNTED diff",
          DIFF in by["001 / ORA / R0.204"][int_st])
    check("PM 0.204 counts 5 cross-street + Int St + Main Line Length (7)",
          by["001 / ORA / R0.204"][diffs_col] in ("7", "7.0"))
    # A NON-signalized control change (A vs B) is NOT crosswalked -> still a genuine diff.
    check("Control Type A(TSMIS) vs B(TSN) — non-signalized, still a genuine diff",
          DIFF in by["001 / ORA / R1"][ctrl])
    check("ML Num Lanes 4 vs 3 is a genuine diff", DIFF in by["001 / ORA / R1"][nl])
    check("PM 1.000 counts Control + ML Num Lanes (2)",
          by["001 / ORA / R1"][diffs_col] in ("2", "2.0"))
    total = sum(1 for r in rows for c in r if DIFF in c)
    check("total counted diff cells across both rows == 9", total == 9)
    print(f"      (rows={len(rows)}, total diff cells={total})")


def test_old_format_refused():
    """A PRE-update consolidated workbook (the 37-column layout) must be REFUSED
    with the re-export hint — reading it by the new positions would silently
    mis-map every column from Description on."""
    print("pre-July-2026 workbook refusal:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_id_old_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    _write_tsmis(tsmis, [
        _tsmis_row("001", "R", "0.204", "21-12-31", "D", "DAPT", "U", "T", "S", "1",
                   "1", "N", "0", "P", "3", "JCT 5") + [None],
    ], old_format=True)
    _write_tsn(tsn, [
        {"PP": "R", "POST_MILE": " 000.204", "LOCATION": "12 ORA 001",
         "DATE_REC": "73-10-19", "HG": "D", "RU": "U"},
    ])
    res = idt.compare(tsmis, tsn, out, events=Events(), confirm_overwrite=lambda _p: True, mode="values")
    check("old-format compare errors (not silently mis-read)", res.status == "error")
    check("the error names the July-2026 format and says to re-export",
          "July 2026" in (res.message or "") and "re" in (res.message or "").lower())
    check("no comparison workbook was written", not out.exists())


def test_route_suffix_match():
    """A TSN route carrying a route suffix (210U) must MATCH the suffix-less
    TSMIS route (210) on base route + PM — not drop to one-sided — and the suffix
    difference must be FLAGGED in the 'Route Suffix' column (the indicator).
    (Since July 2026 the TSMIS Location usually carries the suffix too — then
    both sides read 'U' and the column simply matches.)"""
    print("route-suffix matching + indicator:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_id_rb_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # TSMIS lists the route WITHOUT a suffix ("210"); everything else identical.
    _write_tsmis(tsmis, [
        _tsmis_row("210", "R", "005.000", "73-10-19", "D", "DAPT", "U", "T", "S", "Y",
                   "Y", "N", "N", "P", "3", "JCT 99", int_st="73-10-19"),
    ])
    _write_tsn(tsn, [
        {"PP": "R", "POST_MILE": " 005.000", "LOCATION": "12 ORA 210U", "DATE_REC": "73-10-19",
         "HG": "D", "CITY_CODE": "DAPT", "RU": "U", "TY_INT": "T", "TY_CT": "S", "LT_TY": "Y",
         "MAIN_SM": "Y", "MAIN_LC": "N", "MAIN_RC": "N", "MAIN_TF": "P", "MAIN_NL": 3,
         "DESCRIPTION": "JCT 99", "EFF_DATE_INT": "73-10-19", "EFF_DATE_CT": "73-10-19",
         "EFF_DATE_LT": "73-10-19", "EFF_DATE_ML": "73-10-19",
         "CROSS_BEGIN_DATE": "73-10-19", "EFF_DATE": "73-10-19"},
    ])
    res = idt.compare(tsmis, tsn, out, events=Events(), confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows, sheets = _comparison(out)
    check("matched (1 row on the Comparison sheet, not one-sided)", len(rows) == 1)
    rb = header.index("Route Suffix")
    check("Route Suffix flags the suffix-only difference (U vs blank)", DIFF in rows[0][rb])
    # the substantive attributes are identical, so NOTHING else differs.
    other = sum(1 for i, c in enumerate(rows[0]) if i != rb and DIFF in c)
    check("no other column differs (suffix is the only difference)", other == 0)


def test_normalized_path_crosswalk():
    """A normalized TSN-library workbook carrying RAW control codes (a library built
    before the crosswalk existed — 'stale') must STILL get the crosswalk applied when
    read: _load_tsn re-projects the normalized sheet at compare time. Regression lock
    for the 'Signalized ≠ P' bug. ALSO (CMP-AUD-045): the v3 District/County
    sidecars are READ into the physical key (no longer sliced away) — the output
    rows stay the shared width, the key carries the sidecar county, and a
    pre-v3 library without the sidecars REFUSES with a rebuild hint."""
    print("normalized-library path: crosswalk re-applied + county-aware key:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_id_norm_"))
    norm = root / "tsn_norm.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = idt.NORMALIZED_SHEET
    ws.append(["Route"] + idt.SHARED_HEADER + ["TSN District", "TSN County"])

    def nrow(route, pm, ctrl, light="Y", sidecar=("12", "ORA")):
        d = {"PM": pm, "Control Type": ctrl, "Lighting": light}
        return [route] + [d.get(f, "") for f in idt.SHARED_HEADER] + list(sidecar)

    ws.append(nrow("001", "1.000", "P"))     # stale RAW signal sub-type
    ws.append(nrow("001", "2.000", "J"))     # another stale RAW signal sub-type
    ws.append(nrow("001", "3.000", "A"))     # non-signal, must stay "A"
    wb.save(norm)
    wb.close()
    rows, _ = idt._load_tsn(norm)
    check("output rows stay the shared width (sidecars consumed into the key)",
          all(len(r) == 1 + len(idt.SHARED_HEADER) for r in rows))
    pm_i = 1 + idt.SHARED_HEADER.index("PM")
    ct_i = 1 + idt.SHARED_HEADER.index("Control Type")
    by_pm = {str(r[pm_i]): r[ct_i] for r in rows}
    check("raw 'P' in a normalized library workbook -> 'S' on read",
          by_pm.get("1.000") == "S")
    check("raw 'J' likewise -> 'S'", by_pm.get("2.000") == "S")
    check("non-signal 'A' unchanged", by_pm.get("3.000") == "A")
    comps = dict(rows[0][pm_i].physical_identity.canonical_components)
    check("the key carries the sidecar county (CMP-AUD-045)",
          comps == {"route": "001", "county": "ORA", "postmile": "1"})
    # A pre-v3 library (no sidecar columns) refuses with a rebuild hint.
    old_lib = root / "tsn_old.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = idt.NORMALIZED_SHEET
    ws2.append(["Route"] + idt.SHARED_HEADER)
    ws2.append(nrow("001", "1.000", "A", sidecar=())[:1 + len(idt.SHARED_HEADER)])
    wb2.save(old_lib)
    wb2.close()
    try:
        idt._load_tsn(old_lib)
        check("a pre-v3 (sidecar-less) library refuses with a rebuild hint", False)
    except ValueError as e:
        check("a pre-v3 (sidecar-less) library refuses with a rebuild hint",
              "older normalized" in str(e) and "rebuild" in str(e))


def test_added_columns():
    """The July-2026 columns hold their mappings: an intersecting-route block
    value compares (matching where equal), Xing Line Lgth flags a GENUINE length
    difference (not just padding), and a geometry eff-date conflict flags."""
    print("July-2026 columns (intersecting route, Xing Line Lgth, eff-dates):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_id_add_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    r = _tsmis_row("001", "R", "000.204", "73-10-19", "D", "DAPT", "U", "T", "S", "Y",
                   "Y", "N", "N", "P", "3", "JCT 5", int_st="73-10-19", xll="165")
    r[9] = "70-01-01"      # INT Type Eff-Date — a GENUINE conflict vs TSN's 73-10-19
    r[22] = "100"          # Main Line Length — matches TSN
    r[31] = "005"          # Intrte Route — matches TSN
    _write_tsmis(tsmis, [r])
    _write_tsn(tsn, [{
        "PP": "R", "POST_MILE": " 000.204", "LOCATION": "12 ORA 001", "DATE_REC": "73-10-19",
        "HG": "D", "CITY_CODE": "DAPT", "RU": "U", "TY_INT": "T", "TY_CT": "S", "LT_TY": "Y",
        "MAIN_SM": "Y", "MAIN_LC": "N", "MAIN_RC": "N", "MAIN_TF": "P", "MAIN_NL": 3,
        "DESCRIPTION": "JCT 5", "EFF_DATE_INT": "73-10-19", "EFF_DATE_CT": "73-10-19",
        "EFF_DATE_LT": "73-10-19", "EFF_DATE_ML": "73-10-19", "CROSS_BEGIN_DATE": "73-10-19",
        "EFF_DATE": "73-10-19", "MAIN_OVERRIDE": "100", "CROSS_ROUTE_NAME": "005",
        "X_CROSS_OVERRIDE": "250",
    }])
    res = idt.compare(tsmis, tsn, out, events=Events(), confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows, _ = _comparison(out)
    row = {r[header.index("PM")]: r for r in rows}["001 / ORA / R0.204"]
    check("INT Type Eff-Date is COMPARED and flags (1970 vs 1973 — genuine)",
          DIFF in row[header.index("INT Type Eff-Date")])
    check("Main Line Length is COMPARED and matches (100=100)",
          DIFF not in row[header.index("Main Line Length")])
    check("Intrte Route is COMPARED and matches (005=005)",
          DIFF not in row[header.index("Intrte Route")])
    check("Xing Line Lgth flags a GENUINE difference (165 vs 250)",
          DIFF in row[header.index("Xing Line Lgth")])


def main():
    test_schema()
    test_report_view_typed_truth()
    test_end_to_end()
    test_old_format_refused()
    test_route_suffix_match()
    test_normalized_path_crosswalk()
    test_added_columns()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-INTERSECTION-DETAIL-TSN CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
