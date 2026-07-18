"""Golden check for the shared vs-TSN file-comparator substrate (P5b / S04):
scripts/compare_tsn_common.py + the five thin compare_*_tsn modules that delegate to it.

Two halves:
  * SUBSTRATE — locks compare_tsn_common's exact behavior: the run_files_compare branch
    order + strings (deps gate, per-side missing-file message, the 6-line banner, loader
    ValueError wrap, warnings None -> the run_compare () default), the norm_pm / iso_date
    canon (incl. Intersection Detail's 2-digit TSN year), and that make_notes_writer emits
    one "Notes" sheet with the title + body lines.
  * DELEGATION — proves the five comparators were actually collapsed onto it (this is the
    half that is RED before the refactor): each imports compare_tsn_common and routes
    compare() through run_files_compare; Ramp/Intersection Detail alias _norm_pm/_iso_date
    to the shared helpers; Highway Sequence + Intersection Detail build their Notes legend
    via make_notes_writer. compare_core is never imported here for mutation — untouched.

run_compare is monkeypatched for the happy-path branch (no Excel engine needed); the
make_notes_writer assertion writes a real in-memory workbook. Offline, CI-safe. Run:
    build\\.venv\\Scripts\\python.exe build\\check_compare_tsn_common.py
"""
import contextlib
import inspect
import os
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

import compare_tsn_common as ctc            # noqa: E402
from events import ConsolidateResult, Events  # noqa: E402

_fail = []
_SCRIPTS = ROOT / "scripts"
_MODULES = ["compare_ramp_detail_tsn", "compare_ramp_summary_tsn",
            "compare_highway_sequence_tsn", "compare_intersection_detail_tsn",
            "compare_intersection_summary_tsn"]
_GUARDED_FUNCTION_MODULES = ["compare_highway_log", "compare_highway_detail_tsn",
                             *_MODULES]
_GUARDED_INSTANCE_MODULES = ["compare_highway_detail_pdf",
                             "compare_highway_log_pdf",
                             "compare_highway_sequence_pdf",
                             "compare_intersection_detail_pdf",
                             "compare_ramp_detail_pdf"]


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


@contextlib.contextmanager
def _patch(obj, name, value):
    old = getattr(obj, name)
    setattr(obj, name, value)
    try:
        yield
    finally:
        setattr(obj, name, old)


def _events():
    logs = []
    return Events(on_log=logs.append), logs


def _src(mod):
    return (_SCRIPTS / f"{mod}.py").read_text(encoding="utf-8")


# --------------------------------------------------------------------------- #
# SUBSTRATE
# --------------------------------------------------------------------------- #
def test_normalizers():
    print("compare_tsn_common normalizers:")
    check("norm_pm strips TSN zero-pad (' 000.606' == '0.606')",
          ctc.norm_pm(" 000.606") == ctc.norm_pm("0.606") == "0.606")
    check("norm_pm keeps a leading-dot decimal ('.5' -> '0.5') + sign ('-000.5' -> '-0.5')",
          ctc.norm_pm(".5") == "0.5" and ctc.norm_pm("-000.5") == "-0.5")
    check("norm_pm empty -> ''", ctc.norm_pm(None) == "" and ctc.norm_pm("   ") == "")
    # D1/BUG-07: a real numeric 0 canonicalizes to '0', never blank — this
    # normalizer feeds the row-ALIGNMENT keys, so a blanked 0 mis-aligned rows
    # (the falsy-zero idiom behind v0.18.3's phantom diffs).
    check("norm_pm numeric 0 -> '0' (not blank; alignment-key safe)",
          ctc.norm_pm(0) == "0" and ctc.norm_pm("0") == "0")
    check("iso_date numeric-ish input never blanks via falsy-zero",
          ctc.iso_date(None) == "" and ctc.iso_date(0) != "" or ctc.iso_date(0) == "0")
    check("iso_date TSMIS MM/DD/YYYY -> ISO", ctc.iso_date("02/25/1976") == "1976-02-25")
    check("iso_date TSN 'YYYY-MM-DD HH:MM:SS' -> ISO", ctc.iso_date("1992-09-28 00:00:00") == "1992-09-28")
    check("iso_date TSN 2-digit year windowed (>=30 -> 19xx, <30 -> 20xx)",
          ctc.iso_date("73-10-19") == "1973-10-19" and ctc.iso_date("29-01-02") == "2029-01-02")
    check("iso_date passthrough on unrecognized + empty",
          ctc.iso_date("n/a") == "n/a" and ctc.iso_date(None) == "")
    # CMP-AUD-038: full-match + calendar-aware. Trailing corruption and
    # calendar-impossible dates are PRESERVED as a visible difference, never
    # silently erased or faked into a plausible ISO string. Valid forms (incl.
    # leap days, timestamps, the 2-digit window) keep normalizing unchanged.
    check("iso_date valid leap day 02/29/2000 -> ISO",
          ctc.iso_date("02/29/2000") == "2000-02-29")
    check("iso_date valid ISO timestamp with leap day -> date",
          ctc.iso_date("2000-02-29 12:34:56") == "2000-02-29")
    check("iso_date trailing junk on MM/DD/YYYY preserved (not erased)",
          ctc.iso_date("02/25/1976 junk") == "02/25/1976 junk")
    check("iso_date impossible day 02/31/1976 preserved (not faked)",
          ctc.iso_date("02/31/1976") == "02/31/1976")
    check("iso_date impossible month 13/01/2000 preserved",
          ctc.iso_date("13/01/2000") == "13/01/2000")
    check("iso_date non-leap 02/29/1900 preserved (1900 is not a leap year)",
          ctc.iso_date("02/29/1900") == "02/29/1900")
    check("iso_date trailing junk on ISO date preserved",
          ctc.iso_date("1976-02-25 garbage") == "1976-02-25 garbage")
    check("iso_date impossible 2-digit-year date preserved",
          ctc.iso_date("73-13-45") == "73-13-45")
    check("iso_date impossible ISO timestamp date preserved (not cleaned)",
          ctc.iso_date("1976-02-31 00:00:00") == "1976-02-31 00:00:00")


def test_notes_writer():
    print("compare_tsn_common.make_notes_writer:")
    from openpyxl import Workbook
    writer = ctc.make_notes_writer("My Title", ("line one", "line two"))
    wb = Workbook()
    ws = writer(wb)
    check("creates a sheet named 'Notes'", ws.title == "Notes" and "Notes" in wb.sheetnames)
    _tc = ws.sheet_properties.tabColor
    _rgb = getattr(_tc, "rgb", _tc)              # openpyxl wraps the value in a Color
    check("orange tab color", isinstance(_rgb, str) and _rgb.endswith("ED7D31"))
    col = [r[0].value for r in ws.iter_rows()]
    check("row 1 = title, then one row per body line",
          col == ["My Title", "line one", "line two"])
    check("column A widened", ws.column_dimensions["A"].width == 110)


def test_driver_branches():
    print("run_files_compare branches (deps / missing file / loader error):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_ctc_"))
    a, b, out = root / "a.xlsx", root / "b.xlsx", root / "o.xlsx"
    a.write_bytes(b"x")
    b.write_bytes(b"x")

    def _loader(_t, _n):
        return [["r"]], [["r"]], None

    # deps gate -> the custom deps message, run_compare never reached.
    r = ctc.run_files_compare("SC", a, b, out, banner="B", has_route=True, loader=_loader,
                              deps_ok=False, deps_msg="Required components are missing (pdfplumber, openpyxl).",
                              events=_events()[0])
    check("deps_ok False -> error with the EXACT custom deps message",
          r.status == "error" and r.message == "Required components are missing (pdfplumber, openpyxl).")

    # per-side missing file -> names the side + the path (TSMIS first, TSN second).
    r = ctc.run_files_compare("SC", root / "missing.xlsx", b, out, banner="B",
                              has_route=True, loader=_loader, events=_events()[0])
    check("missing TSMIS -> 'The TSMIS file doesn't exist:\\n<path>'",
          r.status == "error" and r.message == f"The TSMIS file doesn't exist:\n{root / 'missing.xlsx'}")
    r = ctc.run_files_compare("SC", a, root / "missing.xlsx", out, banner="B",
                              has_route=True, loader=_loader, events=_events()[0])
    check("missing TSN -> 'The TSN file doesn't exist:\\n<path>'",
          r.status == "error" and r.message == f"The TSN file doesn't exist:\n{root / 'missing.xlsx'}")

    # loader ValueError -> wrapped to an error result (its message verbatim).
    def _bad(_t, _n):
        raise ValueError("bad shape")

    r = ctc.run_files_compare("SC", a, b, out, banner="B", has_route=True, loader=_bad,
                              events=_events()[0])
    check("loader ValueError -> error result with that message", r.status == "error" and r.message == "bad shape")


def test_driver_happy_path():
    print("run_files_compare happy path (banner + run_compare hand-off):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_ctc_"))
    a, b, out = root / "tsmis.xlsx", root / "tsn.xlsx", root / "o.xlsx"
    a.write_bytes(b"x")
    b.write_bytes(b"x")
    seen = {}

    def _fake_run_compare(sc, rows_t, rows_n, has_route, out_path, **kw):
        seen.update(sc=sc, rows_t=rows_t, rows_n=rows_n, has_route=has_route,
                    out_path=out_path, **kw)
        from openpyxl import Workbook
        wb = Workbook(); wb.active.title = "Comparison"; wb.save(out_path); wb.close()
        return ConsolidateResult(status="ok", output_path=str(out_path))

    ev, logs = _events()
    with _patch(ctc, "run_compare", _fake_run_compare):
        # warnings None from the loader must reach run_compare as the () default.
        ctc.run_files_compare("SCHEMA", a, b, out, banner="Ramp Detail Comparison — TSMIS vs TSN",
                              has_route=True, loader=lambda _t, _n: ([["t"]], [["n"]], None),
                              mode="values", confirm_overwrite=lambda _p: True, events=ev)
    # CMP-AUD-076: two full-selection lines follow the concise-name pair (the
    # basename alone is ambiguous — A\same.xlsx vs B\same.xlsx).
    check("banner == the 8 canonical log lines (incl. the two full selections)",
          logs[:5] == ["=" * 60, "Ramp Detail Comparison — TSMIS vs TSN", "=" * 60,
                       "TSMIS: tsmis.xlsx", "TSN:   tsn.xlsx"]
          and len(logs) == 8 and logs[7] == ""
          and logs[5].startswith("  TSMIS selection: ") and "tsmis.xlsx" in logs[5]
          and "(sha256 " in logs[5]
          and logs[6].startswith("  TSN selection: ") and "tsn.xlsx" in logs[6])
    check("run_compare got schema/has_route/mode/names via a transactional temp path",
          seen.get("sc") == "SCHEMA" and seen.get("has_route") is True
          and Path(seen.get("out_path")).parent == out.parent
          and ".tmp-" in Path(seen.get("out_path")).name
          and seen.get("mode") == "values" and out.is_file()
          and seen.get("name_a") == "tsmis.xlsx" and seen.get("name_b") == "tsn.xlsx"
          and callable(seen.get("confirm_overwrite")))
    check("loader rows reach run_compare in order", seen.get("rows_t") == [["t"]] and seen.get("rows_n") == [["n"]])
    check("warnings None normalized to the run_compare () default", seen.get("warnings") == ())

    # A list of warnings passes straight through (the AGGREGATE summaries' path).
    seen.clear()
    with _patch(ctc, "run_compare", _fake_run_compare):
        ctc.run_files_compare("SCHEMA", a, b, out, banner="B", has_route=False,
                              loader=lambda _t, _n: ([], [], ["a warning"]), events=_events()[0])
    check("explicit warnings list passes through unchanged", seen.get("warnings") == ["a warning"])


def test_driver_target_guard():
    print("run_files_compare target-aware guard reaches transaction + serializer:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_ctc_guard_"))
    a, b, out = root / "a.xlsx", root / "b.xlsx", root / "out.xlsx"
    a.write_bytes(b"a")
    b.write_bytes(b"b")
    guarded = []
    seen = {}

    def guard(path, **binding):
        p = Path(path)
        guarded.append((p, dict(binding)))
        return p == root or p.parent == root

    def _fake_run_compare(_sc, _rows_t, _rows_n, _has_route, out_path, **kw):
        seen.update(out_path=Path(out_path), **kw)
        from openpyxl import Workbook
        wb = Workbook(); wb.active.title = "Comparison"; wb.save(out_path); wb.close()
        return ConsolidateResult(status="ok", output_path=str(out_path))

    with _patch(ctc, "run_compare", _fake_run_compare):
        result = ctc.run_files_compare(
            "SC", a, b, out, banner="B", has_route=True,
            loader=lambda _a, _b: ([["a"]], [["b"]], None),
            mode="values", events=_events()[0], commit_guard=guard)
    guarded_paths = [p for p, _binding in guarded]
    check("guarded driver succeeds and publishes the selected final",
          result.status == "ok" and out.is_file())
    check("the exact final and unpredictable transaction temp are guarded",
          out in guarded_paths and seen.get("out_path") in guarded_paths
          and ".tmp-" in seen.get("out_path").name)
    check("the original target-aware callback reaches compare_core",
          seen.get("commit_guard") is guard)
    check("transaction temp checks carry a bound-parent identity",
          any(p == seen.get("out_path") and binding.get("anchor_identity") is not None
              for p, binding in guarded))

    ran = [False]

    def deny_temp(path, **_binding):
        return ".tmp-" not in Path(path).name

    def _must_not_run(*_args, **_kwargs):
        ran[0] = True
        return ConsolidateResult(status="ok")

    with _patch(ctc, "run_compare", _must_not_run):
        denied = ctc.run_files_compare(
            "SC", a, b, root / "denied.xlsx", banner="B", has_route=True,
            loader=lambda _a, _b: ([["a"]], [["b"]], None),
            mode="values", events=_events()[0], commit_guard=deny_temp)
    check("a guard that denies the allocated temp blocks serializer entry",
          denied.status == "error" and not ran[0]
          and not (root / "denied.xlsx").exists())
    check("denial leaves no transaction temp behind",
          not list(root.glob("*.tmp-*")))


def test_compare_core_guard_boundary():
    print("compare_core checks the exact output again immediately before save:")
    import compare_core
    root = Path(tempfile.mkdtemp(prefix="tsmis_core_guard_"))
    out = root / "core.xlsx"
    schema = compare_core.CompareSchema(
        report_name="Guard Probe", header=["Key", "Value"],
        id_noun="row", id_noun_plural="rows")
    output_checks = [0]

    def revoke_at_save(path):
        p = Path(path)
        if p == out:
            output_checks[0] += 1
            # initial selection + per-flavor entry pass; revoke on the exact
            # third check, immediately before parent.mkdir/wb.save.
            return output_checks[0] < 3
        return True

    result = compare_core.run_compare(
        schema, [["1", "a"]], [["1", "a"]], False, out,
        mode="values", commit_guard=revoke_at_save)
    check("late guard loss returns an ownership error",
          result.status == "error" and "ownership" in result.message.lower())
    check("...no workbook is opened at the denied path", not out.exists())


def test_all_comparator_guard_facades():
    print("every file/PDF comparator exposes and forwards optional commit_guard:")
    for name in _GUARDED_FUNCTION_MODULES:
        mod = __import__(name)
        sig = inspect.signature(mod.compare)
        src = _src(name)
        check(f"{name}: compare(commit_guard=None) is public-compatible",
              "commit_guard" in sig.parameters
              and sig.parameters["commit_guard"].default is None)
        check(f"{name}: forwards the callback to the shared driver",
              "commit_guard=commit_guard" in src)
    for name in _GUARDED_INSTANCE_MODULES:
        mod = __import__(name)
        sig = inspect.signature(mod.TSMIS_PDF_VS_TSN.compare)
        src = _src(name)
        check(f"{name}: adapter.compare(commit_guard=None) is public-compatible",
              "commit_guard" in sig.parameters
              and sig.parameters["commit_guard"].default is None)
        check(f"{name}: forwards the callback to the shared driver",
              "commit_guard=commit_guard" in src)


def test_driver_output_source_aliases():
    """The public FILE driver must be safe without a GUI/Matrix wrapper."""
    print("run_files_compare rejects direct and derived output/source aliases:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_ctc_alias_"))
    a = root / "selected-source.xlsx"
    b = root / "other-source.xlsx"
    a.write_bytes(b"selected source A")
    b.write_bytes(b"selected source B")
    calls = []

    def _destructive_run_compare(_sc, _rows_t, _rows_n, _has_route, out_path, **kw):
        calls.append(Path(out_path))
        Path(out_path).write_bytes(b"comparison replacement")
        if kw.get("mode") == "both":
            p = Path(out_path)
            p.with_name(f"{p.stem} (values){p.suffix}").write_bytes(
                b"derived comparison replacement")

        class _R:
            status = "ok"
        return _R()

    loader = lambda _t, _n: ([['t']], [['n']], None)
    with _patch(ctc, "run_compare", _destructive_run_compare):
        prior = a.read_bytes()
        direct = ctc.run_files_compare(
            "SCHEMA", a, b, a, banner="B", has_route=True, loader=loader,
            mode="values", events=_events()[0])
        check("direct driver rejects a picked output that is source A",
              direct.status == "error" and calls == [])
        check("...source A is byte-for-byte preserved", a.read_bytes() == prior)

        calls.clear()
        picked = root / "comparison.xlsx"
        derived_source = picked.with_name(
            f"{picked.stem} (values){picked.suffix}")
        derived_source.write_bytes(b"selected source used by derived twin")
        prior_derived = derived_source.read_bytes()
        derived = ctc.run_files_compare(
            "SCHEMA", derived_source, b, picked, banner="B", has_route=True,
            loader=loader, mode="both", events=_events()[0])
        check("mode=both rejects an unselected values twin that is source A",
              derived.status == "error" and calls == [])
        check("...the derived-twin source is byte-for-byte preserved",
              derived_source.read_bytes() == prior_derived and not picked.exists())

    # The direct driver now publishes through the same transactional boundary,
    # so an alias that appears only after run_compare has serialized its temp is
    # still rejected immediately before the final replace.
    late_a = root / "late-source.xlsx"
    late_b = root / "late-other.xlsx"
    late_out = root / "late-output.xlsx"
    late_a.write_bytes(b"late source")
    late_b.write_bytes(b"late other")
    late_prior = late_a.read_bytes()
    late_calls = []

    def _late_alias_run(_sc, _rows_t, _rows_n, _has_route, out_path, **_kw):
        from openpyxl import Workbook
        late_calls.append(Path(out_path))
        wb = Workbook(); wb.active.title = "Comparison"; wb.save(out_path); wb.close()
        os.link(late_a, late_out)
        return ConsolidateResult(status="ok", output_path=str(out_path))

    try:
        with _patch(ctc, "run_compare", _late_alias_run):
            late = ctc.run_files_compare(
                "SCHEMA", late_a, late_b, late_out, banner="B", has_route=True,
                loader=loader, mode="values", events=_events()[0])
    except OSError:
        check("late hardlink direct-driver probe skipped only when links are unavailable", True)
    else:
        check("a direct-driver destination that aliases late is rejected after serialization",
              late.status == "error" and len(late_calls) == 1)
        check("...the late-aliased source remains byte-for-byte intact",
              late_a.read_bytes() == late_prior and late_out.read_bytes() == late_prior)


# --------------------------------------------------------------------------- #
# DELEGATION  (the RED-before-refactor half)
# --------------------------------------------------------------------------- #
def test_all_delegate():
    print("every compare_*_tsn delegates to compare_tsn_common:")
    for mod in _MODULES:
        s = _src(mod)
        check(f"{mod}: imports compare_tsn_common", "import compare_tsn_common" in s)
        check(f"{mod}: compare() routes through run_files_compare", "run_files_compare" in s)


def test_detail_aliases():
    print("the two FLAT detail comparators reuse the shared normalizers:")
    import compare_ramp_detail_tsn as rd
    import compare_intersection_detail_tsn as idt
    check("ramp_detail._norm_pm IS compare_tsn_common.norm_pm", rd._norm_pm is ctc.norm_pm)
    check("ramp_detail._iso_date IS compare_tsn_common.iso_date", rd._iso_date is ctc.iso_date)
    check("intersection_detail._norm_pm IS compare_tsn_common.norm_pm", idt._norm_pm is ctc.norm_pm)
    check("intersection_detail._iso_date IS compare_tsn_common.iso_date", idt._iso_date is ctc.iso_date)
    # the canary-pinned behavior still holds through the alias
    check("aliased _norm_pm canon intact (' 000.204' -> '0.204')", rd._norm_pm(" 000.204") == "0.204")
    check("aliased _iso_date 2-digit-year intact ('73-10-19' -> '1973-10-19')",
          idt._iso_date("73-10-19") == "1973-10-19")


def test_notes_delegation():
    print("Highway Sequence builds Notes via make_notes_writer; Intersection Detail has its own:")
    from openpyxl import Workbook
    import compare_highway_sequence_tsn as hs
    import compare_intersection_detail_tsn as idt
    # Highway Sequence still uses the shared (flat) make_notes_writer.
    check("highway_sequence: source uses make_notes_writer",
          "make_notes_writer" in _src(hs.__name__))
    # Intersection Detail OUTGREW the shared writer (v0.17.8 / CR-002 §9e): its Notes sheet
    # has SECTIONS (normalizations applied / columns that differ wholesale / Report View)
    # that the flat make_notes_writer can't express, so it defines its OWN _write_notes_sheet.
    check("intersection_detail: defines its own sectioned _write_notes_sheet (not make_notes_writer)",
          "def _write_notes_sheet" in _src(idt.__name__)
          and "make_notes_writer" not in _src(idt.__name__))
    for mod, label in ((hs, "highway_sequence"), (idt, "intersection_detail")):
        check(f"{label}: schema legend_writer is wired", mod._SCHEMA.legend_writer is not None)
        wb = Workbook()
        mod._SCHEMA.legend_writer(wb)
        check(f"{label}: legend_writer emits a 'Notes' sheet", "Notes" in wb.sheetnames)


def test_same_source_render_equivalence():
    """The same-source (PDF vs Excel) render-artifact rule (owner ruling
    2026-07-16, the CMP-AUD-197 class): OOXML escapes decode, edge tab padding
    never counts, PhysicalKey cells pass through by identity, real data
    differences still flag — and every PDF-vs-Excel flavor (and ONLY those)
    opts in."""
    print("same-source render equivalence (CMP-AUD-197 class):")
    import comparison_contract as cc
    from compare_core import _xl_trim
    t = ctc.same_source_render_text
    # the owner-reported Intersection Detail class: Excel edge tab padding
    check("trailing-tab Excel padding compares equal",
          _xl_trim(t("HILLCREST RD\t\t")) == _xl_trim(t("HILLCREST RD"))
          and _xl_trim(t("HARRIS ROAD \t\t")) == _xl_trim(t("HARRIS ROAD")))
    # OOXML escapes: both hex cases decode; _x005F_ preserves a literal
    check("OOXML _x000d_/_x000D_ escapes decode away at the edge",
          t("CACTUS CITY REST AREA_x000d_") == "CACTUS CITY REST AREA"
          and t("FOO_x000D_") == "FOO")
    check("_x005F_ escaping preserves the literal token (OOXML spec)",
          t("TAG_x005F_x000d_") == "TAG_x000d_")
    check("interior encoded breaks keep separation (a space)",
          _xl_trim(t("LINE1_x000d_LINE2")) == _xl_trim("LINE1 LINE2"))
    check("real data differences still differ",
          _xl_trim(t("HILLCREST RD")) != _xl_trim(t("HILLCREST ROAD")))
    ident = cc.make_physical_identity(
        "001", "ORA", "R1.000", (cc.RawIdentityClaim("route", "001"),), "x")
    key = cc.physical_key("R1.000", ident)
    rows = ctc.same_source_render_rows([["001", "ORA", key, "DESC\t"]])
    check("PhysicalKey cells pass through by identity; text normalizes",
          rows[0][2] is key and type(rows[0][2]) is cc.PhysicalKey
          and rows[0][3] == "DESC")
    # the flavors: every PDF-vs-Excel self-check opts in; no vs-TSN leg does
    import compare_intersection_detail_pdf as idp
    import compare_ramp_detail_pdf as rdp
    import compare_highway_sequence_pdf as hsp
    check("all three PDF-vs-Excel flavors are same-source",
          idp.TSMIS_PDF_VS_EXCEL._same_source
          and rdp.TSMIS_PDF_VS_EXCEL._same_source
          and hsp.TSMIS_PDF_VS_EXCEL._same_source)
    check("no vs-TSN flavor is same-source (oracle byte semantics kept)",
          not idp.TSMIS_PDF_VS_TSN._same_source
          and not rdp.TSMIS_PDF_VS_TSN._same_source
          and not hsp.TSMIS_PDF_VS_TSN._same_source)


def test_decode_ooxml_escapes():
    """The shared decode seam (CMP-AUD-197's HSL vs-TSN half rides it too):
    byte-equivalent to openpyxl.utils.escape.unescape — the Stage-8 oracles'
    xlsx reading — on every censused shape."""
    print("decode_ooxml_escapes matches openpyxl's unescape byte-for-byte:")
    from openpyxl.utils.escape import unescape as xlsx_unescape
    d = ctc.decode_ooxml_escapes
    fixtures = (
        "A_x000d_B",                # the censused lowercase CR escape
        "A_x000D_B",                # hex-digit case-insensitivity
        "_X000D_",                  # uppercase prefix is NOT an escape
        "TAG_x005F_x000d_",         # literal-preserving _x005F_ consumption
        "_x00zz_",                  # non-hex stays literal
        "plain text, no escapes",
        "edge_x0009_",              # encoded tab
        "_x000a__x000d_",           # adjacent escapes
    )
    for fixture in fixtures:
        check(f"decode == openpyxl unescape for {fixture!r}",
              d(fixture) == xlsx_unescape(fixture))
    check("the censused CR decodes to a real CR",
          d("A_x000d_B") == "A\rB")


def main():
    test_normalizers()
    test_notes_writer()
    test_driver_branches()
    test_driver_happy_path()
    test_driver_target_guard()
    test_compare_core_guard_boundary()
    test_driver_output_source_aliases()
    test_all_comparator_guard_facades()
    test_all_delegate()
    test_detail_aliases()
    test_notes_delegation()
    test_same_source_render_equivalence()
    test_decode_ooxml_escapes()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-TSN-COMMON CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
