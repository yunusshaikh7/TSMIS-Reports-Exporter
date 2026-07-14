#!/usr/bin/env python3
"""Highway Sequence Stage-8 source-core comparison checkpoint.

This checkpoint reparses the immutable private TSMIS Excel/PDF capture and all
twelve authoritative TSN PDFs.  It reads the accepted normalized TSN workbook
and accepted Stage-6 result only after binding their exact identities.  It does
not consume a development row cache or draft result, and it never imports a
product parser, comparator, schema, or evidence adapter.

The artifact emitted here is deliberately *not* final Stage-8 acceptance.
Product-publication, workbook/evidence, permanent-gate, detached-decision, and
byte-identical replay proofs are separate layers owned by the final driver.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor
from dataclasses import asdict, dataclass, replace
from functools import lru_cache
import hashlib
import importlib.metadata
import io
import json
import os
from pathlib import Path
import re
import sys
import tempfile
from typing import Callable, Iterable, Mapping, Sequence

import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils.escape import unescape as xlsx_unescape


BUILD_ROOT = Path(__file__).resolve().parent
REPO_ROOT = BUILD_ROOT.parent
sys.path.insert(0, str(BUILD_ROOT))

# Audit-owned dependencies only.  Their exact bytes are bound below before and
# after every execution.
import phase3_xlsx_stream as xlsx_stream  # noqa: E402
import phase6_highway_sequence_conservation as stage6  # noqa: E402
import phase8_highway_sequence_source_oracle as tsmis_source  # noqa: E402


VISUAL_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd"
)
PRIVATE_ROOT = VISUAL_ROOT / "phase8_highway_sequence_private_sources_r1"
NORMALIZED_TSN = (
    VISUAL_ROOT / "phase4_tsn_rebaseline" / "raw-2026-07-12-r7"
    / "highway_sequence" / "consolidated"
    / "tsn_highway_sequence_normalized.xlsx"
)
STAGE6_RESULT = (
    VISUAL_ROOT / "phase6_tsn_conservation"
    / "highway_sequence_conservation_r7.json"
)
STAGE6_ACCEPTANCE = Path(str(STAGE6_RESULT) + ".acceptance.json")
PRODUCT_COMPARATOR = REPO_ROOT / "scripts" / "compare_highway_sequence_tsn.py"
RESULT_NAME = "source-core.json"
INCOMPLETE_R1_ROOT = VISUAL_ROOT / "phase8_highway_sequence_source_core_checkpoint_r1"

STATIC_FILE_BINDINGS = {
    "source_oracle": {
        "path": BUILD_ROOT / "phase8_highway_sequence_source_oracle.py",
        "bytes": 35_385,
        "sha256": "fc4038f96d09c93021919262cc0cf902ae13b962100836d4efe7f016b4df70b3",
    },
    "stage6_oracle": {
        "path": BUILD_ROOT / "phase6_highway_sequence_conservation.py",
        "bytes": 63_233,
        "sha256": "0d6cacfa5a4615a80381b077780b051127958bbf325979cf24b7a5c29eb8e17b",
    },
    "xlsx_reader": {
        "path": BUILD_ROOT / "phase3_xlsx_stream.py",
        "bytes": 40_888,
        "sha256": "bbfda5ccdbea3697978c0ba4414b7dccf3d5c248ba6762aa946c76e920fc940b",
    },
    "accepted_normalized_tsn": {
        "path": NORMALIZED_TSN,
        "bytes": 2_536_901,
        "sha256": "9dc84c661a9284131baf928767e210a6d708c0a338819fca2b69b907f85dd041",
    },
    "accepted_stage6_result": {
        "path": STAGE6_RESULT,
        "bytes": 1_276_684,
        "sha256": "bdd344258ced0e138196c518be2d49ee058f5f9c0f52dea860c328fc3216d1e2",
    },
    "accepted_stage6_decision": {
        "path": STAGE6_ACCEPTANCE,
        "bytes": 5_934,
        "sha256": "71fe59a5f4676d3b935bcbea380374b14fdccfd77b674ea88148fa18760ffde2",
    },
    "modeled_product_comparator": {
        "path": PRODUCT_COMPARATOR,
        "bytes": 12_464,
        "sha256": "08ae1592a060ca8b6be9bf5d6521629c66460e6d4b5381fbde3425cffaeaea03",
    },
}

PACKAGE_BINDINGS = {"openpyxl": "3.1.5", "pdfplumber": "0.11.9"}

VALUE_FIELDS = ("City", "HG", "FT", "Distance To Next Point", "Description")
SAME_SOURCE_FIELDS = (
    "PM Suffix", "City", "HG", "FT", "Distance To Next Point", "Description",
)
ASSERTED_VS_TSN_FIELDS = ("FT", "Description")
ROUTE_PREFIX_RE = re.compile(r"^(\d{1,3}[A-Z]?)/")
PREFIX_WITH_PADDING_RE = re.compile(r"^(\d{1,3}[A-Z]?)/(\s+)")
PM_RE = re.compile(r"^([A-Z]?\d{3}\.\d{3})(E?)$")

EXPECTED_PREFIX_LEDGER = {
    "wire_bytes": 7_517,
    "sha256": "5dacffd43c62ea8001796e5b4d87d1290b07cd7084861f26cf8cf047d452eab7",
    "content_sorted_sha256": "59f3afe3336d07daaf5fd6e228b060ab5e822c1040f2e21e3dd2fca88b9d11e7",
}
EXPECTED_UNKNOWN_RAW_PUBLICATION_SHA256 = (
    "bbd85ad3b3de2bf5312e6a2945270b4d1a521acc690de62dabe833a810f8aeab"
)
EXPECTED_PADDING_KEYS = (
    ("005", "ORA", "020.746"),
    ("070", "YUB", "000.204"),
    ("073", "ORA", "016.689"),
)
EXPECTED_COLLAPSED_DUPLICATE_KEYS = (
    ("028", "PLA", "009.880"),
    ("145", "FRE", "033.129"),
)

# Source-owned final design table.  The four TSN legs retain all context-field
# truth and separately identify the currently asserted FT/Description surface.
EXPECTED_CURRENT_LEGS = {
    "excel_vs_raw_tsn": {
        "shape": (60_494, 69_804, 57_072, 3_422, 12_732),
        "all": (23_691, 30_005, {
            "City": 15_026, "Description": 4_894,
            "Distance To Next Point": 6_972, "FT": 695, "HG": 2_418,
        }),
        "asserted": (4_894, 5_589, {"Description": 4_894, "FT": 695}),
    },
    "excel_vs_normalized_tsn": {
        "shape": (60_494, 69_758, 57_072, 3_422, 12_686),
        "all": (23_692, 30_277, {
            "City": 15_023, "Description": 4_894,
            "Distance To Next Point": 7_243, "FT": 700, "HG": 2_417,
        }),
        "asserted": (4_895, 5_594, {"Description": 4_894, "FT": 700}),
    },
    "pdf_vs_raw_tsn": {
        "shape": (60_493, 69_804, 57_505, 2_988, 12_299),
        "all": (23_872, 29_189, {
            "City": 15_140, "Description": 4_916,
            "Distance To Next Point": 7_056, "FT": 85, "HG": 1_992,
        }),
        "asserted": (4_916, 5_001, {"Description": 4_916, "FT": 85}),
    },
    "pdf_vs_normalized_tsn": {
        "shape": (60_493, 69_758, 57_505, 2_988, 12_253),
        "all": (23_873, 29_461, {
            "City": 15_137, "Description": 4_916,
            "Distance To Next Point": 7_327, "FT": 90, "HG": 1_991,
        }),
        "asserted": (4_917, 5_006, {"Description": 4_916, "FT": 90}),
    },
    "pdf_vs_excel": {
        "shape": (60_493, 60_494, 60_493, 0, 1),
        "all": (1_410, 3_721, {
            "Description": 1_133, "FT": 1_129, "HG": 910, "PM Suffix": 549,
        }),
        "asserted": (1_410, 3_721, {
            "Description": 1_133, "FT": 1_129, "HG": 910, "PM Suffix": 549,
        }),
    },
}

EXPECTED_HISTORICAL_LEGS = {
    "historical_excel_vs_raw_tsn": {
        "shape": (60_493, 69_804, 57_071, 3_422, 12_733),
        "all": (23_695, 30_009, {
            "City": 15_026, "Description": 4_898,
            "Distance To Next Point": 6_972, "FT": 695, "HG": 2_418,
        }),
        "asserted": (4_898, 5_593, {"Description": 4_898, "FT": 695}),
        "assignment_cost": (30_009, 157_065, 1_856),
        "duplicate_trace_sha256": "896f2be36d8f2f5474331c0f79afa87ca4fc3c02df55f7e6f332ca7ca534538e",
    },
    "historical_excel_vs_normalized_tsn": {
        "shape": (60_493, 69_758, 57_071, 3_422, 12_687),
        "all": (23_696, 30_281, {
            "City": 15_023, "Description": 4_898,
            "Distance To Next Point": 7_243, "FT": 700, "HG": 2_417,
        }),
        "asserted": (4_899, 5_598, {"Description": 4_898, "FT": 700}),
        "assignment_cost": (30_281, 159_641, 1_850),
        "duplicate_trace_sha256": "db5d93c9d52e718118c07da770545ec23b22f10a65fc2867a182ff384ae3e829",
    },
}


class SourceCoreError(RuntimeError):
    """An immutable binding or source-owned semantic invariant failed."""


@dataclass(frozen=True)
class Row:
    dataset: str
    source_role: str
    source_format: str
    record_kind: str
    source_ref: str
    source_index: int
    route: str
    county: str
    pm_base: str
    pm_suffix: str
    values: tuple[str, ...]
    raw_values: tuple[object, ...]
    provenance: dict[str, object]

    @property
    def pm_full(self) -> str:
        return self.pm_base + self.pm_suffix

    @property
    def identity(self) -> tuple[str, str, str]:
        return self.route, self.county, self.pm_full


@dataclass(frozen=True)
class CapturedMember:
    name: str
    bytes: int
    sha256: str
    payload: bytes


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise SourceCoreError(message)


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_bytes(value: object, *, newline: bool = True) -> bytes:
    text = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
    )
    return (text + ("\n" if newline else "")).encode("utf-8")


def _identity(path: Path) -> dict[str, object]:
    path = path.resolve()
    return {
        "canonical_path": str(path),
        "bytes": path.stat().st_size,
        "sha256": _sha_file(path),
    }


def _bind_static_inputs() -> dict[str, object]:
    files: dict[str, object] = {}
    for label, spec in STATIC_FILE_BINDINGS.items():
        observed = _identity(Path(spec["path"]))
        _require(
            observed["bytes"] == spec["bytes"]
            and observed["sha256"] == spec["sha256"],
            f"{label} identity drift: {observed}",
        )
        files[label] = observed
    packages = {}
    for distribution, expected_version in PACKAGE_BINDINGS.items():
        observed_version = importlib.metadata.version(distribution)
        _require(
            observed_version == expected_version,
            f"{distribution} version drift: {observed_version}",
        )
        module = sys.modules[distribution]
        packages[distribution] = {
            "version": observed_version,
            "module_identity": _identity(Path(module.__file__)),
        }
    return {
        "self": _identity(Path(__file__)),
        "files": files,
        "packages": packages,
        "capture": tsmis_source._bind_capture(),
    }


def _capture_tree_payloads(
    label: str, suffix: str, capture_binding: Mapping[str, object],
) -> tuple[list[CapturedMember], dict[str, object]]:
    root = PRIVATE_ROOT / label
    expected_entries = {
        item["name"]: item for item in capture_binding["trees"][label]["members"]
    }
    paths = sorted(root.glob(f"*{suffix}"), key=lambda path: path.name)
    _require(
        [path.name for path in paths] == sorted(expected_entries),
        f"{label}: captured member-name universe drift",
    )
    captured: list[CapturedMember] = []
    for path in paths:
        payload = path.read_bytes()
        member = CapturedMember(
            name=path.name,
            bytes=len(payload),
            sha256=_sha_bytes(payload),
            payload=payload,
        )
        expected = expected_entries[path.name]
        _require(
            member.bytes == expected["bytes"]
            and member.sha256 == expected["sha256"],
            f"{label}/{path.name}: single byte-capture identity drift",
        )
        captured.append(member)
    identity_records = [
        {"name": member.name, "bytes": member.bytes, "sha256": member.sha256}
        for member in captured
    ]
    return captured, {
        "members": len(captured),
        "bytes": sum(member.bytes for member in captured),
        "ordered_identity_sha256": _sha_bytes(_json_bytes(identity_records)),
        "identities": identity_records,
        "parse_uses_this_exact_payload_capture": True,
    }


def _excel_member_bytes(
    member: CapturedMember, source: str,
) -> tuple[list[tsmis_source.SourceRow], dict[str, object]]:
    route = tsmis_source._route_from_member(Path(member.name))
    workbook = load_workbook(
        io.BytesIO(member.payload), read_only=True, data_only=False, keep_links=False,
    )
    try:
        _require(
            workbook.sheetnames == [tsmis_source.SHEET_NAME],
            f"{member.name}: sheet role universe {workbook.sheetnames}",
        )
        sheet = workbook[tsmis_source.SHEET_NAME]
        _require(sheet.sheet_state == "visible", f"{member.name}: hidden source sheet")
        _require(
            sheet.max_column == len(tsmis_source.HEADERS),
            f"{member.name}: max-column drift {sheet.max_column}",
        )
        header_cells = next(sheet.iter_rows(
            min_row=1, max_row=1, max_col=len(tsmis_source.HEADERS),
        ))
        header = tuple(cell.value for cell in header_cells)
        _require(
            header == tsmis_source.HEADERS,
            f"{member.name}: positional header drift {header!r}",
        )
        rows: list[tsmis_source.SourceRow] = []
        scalar_types: Counter[str] = Counter()
        for source_row, cells in enumerate(sheet.iter_rows(
            min_row=2, max_col=len(tsmis_source.HEADERS),
        ), 2):
            values: list[str | None] = []
            for cell in cells:
                _require(
                    cell.data_type not in ("f", "e"),
                    f"{member.name}!{cell.coordinate}: formula/error source cell",
                )
                value = cell.value
                if value is None:
                    scalar_types["null"] += 1
                    values.append(None)
                elif isinstance(value, str):
                    scalar_types["str"] += 1
                    values.append(value)
                else:
                    raise SourceCoreError(
                        f"{member.name}!{cell.coordinate}: unsupported source scalar "
                        f"{type(value).__name__}={value!r}"
                    )
            _require(
                not all(value is None for value in values),
                f"{member.name}: blank physical data row {source_row}",
            )
            _require(
                not _text(values[3])
                or tsmis_source.PM_RE.fullmatch(_text(values[3])) is not None,
                f"{member.name}: invalid PM at row {source_row}",
            )
            prefix = _text(values[2])
            suffix = _text(values[4])
            _require(
                not prefix or prefix in tsmis_source.PREFIX_SET,
                f"{member.name}: invalid prefix {prefix!r}",
            )
            _require(
                suffix in ("", "E"),
                f"{member.name}: invalid suffix {suffix!r}",
            )
            rows.append(tsmis_source.SourceRow(
                source=source,
                member=member.name,
                route=route,
                source_index=source_row,
                source_ref=f"{member.name}:row:{source_row}",
                page=None,
                top=None,
                values=tuple(values),
            ))
        _require(
            sheet.max_row == len(rows) + 1,
            f"{member.name}: physical row extent drift",
        )
        return rows, {
            "member": member.name,
            "captured_bytes": member.bytes,
            "captured_sha256": member.sha256,
            "route": route,
            "rows": len(rows),
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "scalar_types": dict(sorted(scalar_types.items())),
        }
    finally:
        workbook.close()


def _parse_excel_captures(
    members: Sequence[CapturedMember], source: str,
) -> tuple[list[tsmis_source.SourceRow], dict[str, object]]:
    rows: list[tsmis_source.SourceRow] = []
    diagnostics = []
    for member in members:
        member_rows, diagnostic = _excel_member_bytes(member, source)
        rows.extend(member_rows)
        diagnostics.append(diagnostic)
    return rows, {
        "members": len(diagnostics),
        "member_diagnostics": diagnostics,
        "all_cells_string_or_null": all(
            set(item["scalar_types"]) <= {"str", "null"}
            for item in diagnostics
        ),
        "parsed_exact_captured_payloads": True,
    }


def _pdf_member_bytes(
    item: tuple[str, bytes, int, str, str],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    name, payload, captured_bytes, captured_sha256, source = item
    route = tsmis_source._route_from_member(Path(name))
    rows: list[tsmis_source.SourceRow] = []
    page_diagnostics = []
    data_pages = 0
    trailer_seen = False
    fragment_count = 0
    with pdfplumber.open(io.BytesIO(payload)) as pdf:
        if len(pdf.pages) < 3:
            raise SourceCoreError(f"{name}: expected cover, legend, and data pages")
        cover = " ".join((pdf.pages[0].extract_text() or "").split())
        legend = " ".join((pdf.pages[1].extract_text() or "").split())
        cover_claims = (
            "California Department of Transportation", "Highway Sequence Listing",
            "REPORT DATE : 7/9/2026", "REFERENCE DATE : 7/9/2026",
            f"ROUTE : {route}", "TASAS \u2013 TSMIS",
        )
        missing_cover = [claim for claim in cover_claims if claim not in cover]
        if missing_cover:
            raise SourceCoreError(f"{name}: cover claims missing {missing_cover}")
        legend_claims = (
            "Legend", "Route Suffix Codes", "Post Mile Prefix Codes",
            "Post Mile Suffix Codes", "E - Equation", "Font Color",
        )
        missing_legend = [claim for claim in legend_claims if claim not in legend]
        if missing_legend:
            raise SourceCoreError(f"{name}: legend claims missing {missing_legend}")

        for page_number, page in enumerate(pdf.pages[2:], 3):
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            lines = tsmis_source._cluster_lines(words)
            try:
                header = tsmis_source._header_line(lines)
            except tsmis_source.OracleError as error:
                raise SourceCoreError(f"{name} p{page_number}: {error}") from error
            if header is None:
                raise SourceCoreError(f"{name}: page {page_number} has no data header")
            data_pages += 1
            header_top, anchors = header
            header_bottom = max(float(word["bottom"]) for word in anchors.values())
            banner_text = " ".join(
                tsmis_source._line_text(line_words)
                for top, line_words in lines if top < header_top
            )
            route_claims = [
                tsmis_source._route_token(value)
                for value in tsmis_source.ROUTE_CLAIM_RE.findall(banner_text)
            ]
            if route_claims != [route]:
                raise SourceCoreError(
                    f"{name} p{page_number}: route claims {route_claims}"
                )
            directions = [
                f"{left}-{right}"
                for left, right in tsmis_source.DIRECTION_RE.findall(banner_text)
            ]
            if len(directions) != 1:
                raise SourceCoreError(
                    f"{name} p{page_number}: direction claims {directions}"
                )
            districts = [
                value.zfill(2)
                for value in tsmis_source.DISTRICT_RE.findall(banner_text)
            ]
            if len(districts) > 1:
                raise SourceCoreError(
                    f"{name} p{page_number}: district claims {districts}"
                )

            page_rows: list[list[object]] = []
            fragments: list[tuple[float, str]] = []
            unclassified = []
            for top, line_words in lines:
                if top <= header_bottom + 2:
                    continue
                raw_text = tsmis_source._line_text(line_words)
                if raw_text.startswith(tsmis_source.TRAILER_HEADING):
                    trailer_seen = True
                    break
                columns = tsmis_source._fixed_columns(line_words)
                pm = columns["PM"]
                if tsmis_source.PM_RE.fullmatch(pm) or tsmis_source._pmless(columns):
                    prefix = columns["PM Prefix"]
                    suffix = columns["PM Suffix"]
                    if prefix and prefix not in tsmis_source.PREFIX_SET:
                        raise SourceCoreError(
                            f"{name} p{page_number}: unknown prefix {prefix!r}"
                        )
                    if suffix not in ("", "E"):
                        raise SourceCoreError(
                            f"{name} p{page_number}: unknown suffix {suffix!r}"
                        )
                    if columns["HG"] and columns["HG"] not in tsmis_source.HG_SET:
                        raise SourceCoreError(
                            f"{name} p{page_number}: unknown HG {columns['HG']!r}"
                        )
                    if columns["FT"] and columns["FT"] not in tsmis_source.FT_SET:
                        raise SourceCoreError(
                            f"{name} p{page_number}: unknown FT {columns['FT']!r}"
                        )
                    values = [
                        columns[field] or None for field in tsmis_source.FIELD_NAMES
                    ]
                    page_rows.append([top, values, raw_text])
                elif columns["Description"] and not any(
                    columns[field]
                    for field in tsmis_source.FIELD_NAMES if field != "Description"
                ):
                    fragments.append((top, columns["Description"]))
                else:
                    unclassified.append({
                        "top": f"{top:.3f}", "text": raw_text, "columns": columns,
                    })
            if unclassified:
                raise SourceCoreError(
                    f"{name} p{page_number}: unclassified lines {unclassified[:3]}"
                )
            parts: dict[int, list[tuple[float, str]]] = {
                index: [(float(row[0]), _text(row[1][8]))]
                for index, row in enumerate(page_rows)
            }
            for fragment_top, fragment_text in fragments:
                distances = sorted(
                    (abs(float(row[0]) - fragment_top), index)
                    for index, row in enumerate(page_rows)
                )
                if not distances or distances[0][0] > tsmis_source.FRAGMENT_MAX_DISTANCE:
                    raise SourceCoreError(
                        f"{name} p{page_number}: orphan fragment {fragment_text!r}"
                    )
                if len(distances) > 1 and abs(
                    distances[0][0] - distances[1][0]
                ) < 0.001:
                    raise SourceCoreError(
                        f"{name} p{page_number}: ambiguous fragment {fragment_text!r}"
                    )
                parts[distances[0][1]].append((fragment_top, fragment_text))
                fragment_count += 1
            for index, (top, values, _raw_text) in enumerate(page_rows):
                values[8] = tsmis_source._join_description([
                    text for _part_top, text in sorted(parts[index]) if text
                ]) or None
                source_index = len(rows) + 1
                rows.append(tsmis_source.SourceRow(
                    source=source, member=name, route=route,
                    source_index=source_index,
                    source_ref=f"{name}:page:{page_number}:row:{index + 1}",
                    page=page_number, top=f"{float(top):.3f}", values=tuple(values),
                ))
            page_diagnostics.append({
                "page": page_number,
                "rows": len(page_rows),
                "fragments": len(fragments),
                "direction": directions[0],
                "district": districts[0] if districts else None,
                "header_x0": {
                    token: f"{float(word['x0']):.3f}"
                    for token, word in anchors.items()
                },
            })
        metadata = {
            str(key): str(value)
            for key, value in sorted((pdf.metadata or {}).items())
        }
    return [asdict(row) for row in rows], {
        "member": name,
        "captured_bytes": captured_bytes,
        "captured_sha256": captured_sha256,
        "route": route,
        "pages": data_pages + 2,
        "data_pages": data_pages,
        "rows": len(rows),
        "fragments": fragment_count,
        "trailer_seen": trailer_seen,
        "metadata": metadata,
        "page_diagnostics": page_diagnostics,
    }


def _parse_pdf_captures(
    members: Sequence[CapturedMember], workers: int, source: str,
) -> tuple[list[tsmis_source.SourceRow], dict[str, object]]:
    arguments = [
        (member.name, member.payload, member.bytes, member.sha256, source)
        for member in members
    ]
    with ProcessPoolExecutor(max_workers=workers) as executor:
        results = list(executor.map(_pdf_member_bytes, arguments, chunksize=1))
    rows: list[tsmis_source.SourceRow] = []
    diagnostics = []
    for serialized_rows, diagnostic in results:
        rows.extend(tsmis_source.SourceRow(**{
            **row, "values": tuple(row["values"]),
        }) for row in serialized_rows)
        diagnostics.append(diagnostic)
    anchor_values: dict[str, list[float]] = defaultdict(list)
    for member in diagnostics:
        for page in member["page_diagnostics"]:
            for token, value in page["header_x0"].items():
                anchor_values[token].append(float(value))
    return rows, {
        "members": len(diagnostics),
        "pages": sum(item["pages"] for item in diagnostics),
        "data_pages": sum(item["data_pages"] for item in diagnostics),
        "description_fragments": sum(item["fragments"] for item in diagnostics),
        "members_with_trailer": sum(
            bool(item["trailer_seen"]) for item in diagnostics
        ),
        "header_anchor_x0_ranges": {
            token: {"min": f"{min(values):.3f}", "max": f"{max(values):.3f}"}
            for token, values in sorted(anchor_values.items())
        },
        "member_diagnostics": diagnostics,
        "parsed_exact_captured_payloads": True,
    }


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _space(value: object, *, xlsx: bool = False) -> str:
    text = _text(value)
    if xlsx:
        text = xlsx_unescape(text)
    return " ".join(text.split())


def _route(value: object) -> str:
    match = re.fullmatch(r"(\d{1,3})([A-Za-z]?)", _text(value))
    if match is None:
        raise SourceCoreError(f"invalid Route: {value!r}")
    return match.group(1).zfill(3) + match.group(2).upper()


def _county(value: object) -> str:
    return _text(value).rstrip(".").upper()


def _split_pm(value: object) -> tuple[str, str]:
    text = _text(value).upper()
    if not text:
        return "", ""
    match = PM_RE.fullmatch(text)
    if match is None:
        raise SourceCoreError(f"invalid complete postmile: {value!r}")
    return match.group(1), match.group(2)


def _semantic_description(
    value: object,
    route: str,
    *,
    xlsx: bool = False,
    strip_tsmis_route_label: bool = False,
) -> str:
    text = _space(value, xlsx=xlsx)
    if strip_tsmis_route_label:
        match = ROUTE_PREFIX_RE.match(text)
        if match is not None and _route(match.group(1)) == route:
            text = text[match.end():].lstrip()
    return text


def _tsmis_rows(
    source_rows: Sequence[tsmis_source.SourceRow], dataset: str,
) -> list[Row]:
    is_xlsx = "excel" in dataset
    rows: list[Row] = []
    for source in source_rows:
        raw = tuple(source.values)
        route = _route(source.route)
        pm_base = (_text(raw[2]) + _text(raw[3])).upper()
        if pm_base:
            _split_pm(pm_base)
        suffix = _text(raw[4]).upper()
        _require(suffix in ("", "E"), f"{source.source_ref}: invalid suffix")
        rows.append(Row(
            dataset=dataset,
            source_role="tsmis",
            source_format="xlsx" if is_xlsx else "pdf",
            record_kind="tsmis",
            source_ref=source.source_ref,
            source_index=source.source_index,
            route=route,
            county=_county(raw[0]),
            pm_base=pm_base,
            pm_suffix=suffix,
            values=(
                _space(raw[1], xlsx=is_xlsx),
                _space(raw[5], xlsx=is_xlsx),
                _space(raw[6], xlsx=is_xlsx),
                _space(raw[7], xlsx=is_xlsx),
                _semantic_description(raw[8], route, xlsx=is_xlsx),
            ),
            raw_values=raw,
            provenance={
                "member": source.member,
                "page": source.page,
                "top": source.top,
            },
        ))
    return rows


def _raw_tsn_rows(
    records: Sequence[Mapping[str, object]],
) -> tuple[list[Row], list[Row]]:
    known: list[Row] = []
    unknown: list[Row] = []
    for ordinal, item in enumerate(records, 1):
        route = _route(item["route"])
        pm_base, suffix = _split_pm(item["pm"])
        raw = (
            item.get("city"), item.get("hg"), item.get("ft"),
            item.get("distance"), item.get("description"),
        )
        row = Row(
            dataset="raw_tsn",
            source_role="tsn",
            source_format="pdf",
            record_kind=str(item["kind"]),
            source_ref=(
                f"{item['member']}:page:{item['physical_page']}:line:{item['line']}"
            ),
            source_index=ordinal,
            route=route,
            county=_county(item.get("county")),
            pm_base=pm_base,
            pm_suffix=suffix,
            values=(
                _space(item.get("city")),
                _space(item.get("hg")),
                _space(item.get("ft")),
                _space(item.get("distance")),
                _semantic_description(item.get("description"), route),
            ),
            raw_values=raw,
            provenance={
                key: item[key] for key in (
                    "member", "district", "direction", "physical_page",
                    "printed_page", "line", "top", "raw_text",
                )
            },
        )
        (known if row.county else unknown).append(row)
    return known, unknown


def _normalized_tsn_rows(sheet_rows: Sequence[object]) -> list[Row]:
    rows: list[Row] = []
    for item in sheet_rows:
        raw = tuple(item.values)
        _require(len(raw) == 8, f"normalized row {item.source_row}: width drift")
        route = _route(raw[0])
        pm_base, suffix = _split_pm(raw[2])
        description = _space(raw[7])
        rows.append(Row(
            dataset="normalized_tsn",
            source_role="tsn",
            source_format="xlsx",
            record_kind=("equate" if description.startswith("EQUATES TO") else "data"),
            source_ref=f"normalized:row:{item.source_row}",
            source_index=int(item.source_row),
            route=route,
            county=_county(raw[1]),
            pm_base=pm_base,
            pm_suffix=suffix,
            values=(
                _space(raw[3]), _space(raw[4]), _space(raw[5]),
                _space(raw[6]), description,
            ),
            raw_values=raw,
            provenance={"normalized_source_row": item.source_row},
        ))
    return rows


def _same_source_values(row: Row) -> tuple[str, ...]:
    return (row.pm_suffix, *row.values)


def _vs_tsn_values(row: Row) -> tuple[str, ...]:
    values = list(row.values)
    if row.source_role == "tsmis":
        values[-1] = _semantic_description(
            row.raw_values[8],
            row.route,
            xlsx=row.source_format == "xlsx",
            strip_tsmis_route_label=True,
        )
    return tuple(values)


def _char_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    return _char_distance_unequal(left, right)


@lru_cache(maxsize=None)
def _char_distance_unequal(left: str, right: str) -> int:
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for left_index, left_character in enumerate(left, 1):
        current = [left_index]
        for right_index, right_character in enumerate(right, 1):
            current.append(min(
                current[-1] + 1,
                previous[right_index] + 1,
                previous[right_index - 1]
                + (left_character != right_character),
            ))
        previous = current
    return previous[-1]


def _cost(
    left_values: Sequence[str],
    right_values: Sequence[str],
    left_position: int,
    right_position: int,
) -> tuple[int, int, int]:
    return (
        sum(a != b for a, b in zip(left_values, right_values, strict=True)),
        sum(
            _char_distance(a, b)
            for a, b in zip(left_values, right_values, strict=True)
        ),
        abs(left_position - right_position),
    )


def _add_cost(
    left: tuple[int, int, int], right: tuple[int, int, int],
) -> tuple[int, int, int]:
    return tuple(
        a + b for a, b in zip(left, right, strict=True)
    )


def _assign_group(
    left: Sequence[Row],
    right: Sequence[Row],
    projector: Callable[[Row], tuple[str, ...]],
) -> tuple[
    list[tuple[Row, Row]], list[Row], list[Row], tuple[int, int, int],
]:
    if not left or not right:
        return [], list(left), list(right), (0, 0, 0)
    swapped = len(left) > len(right)
    small = right if swapped else left
    large = left if swapped else right
    _require(
        len(large) <= 12,
        f"assignment group too large for exact DP: {len(left)}x{len(right)}",
    )
    small_values = [projector(row) for row in small]
    large_values = [projector(row) for row in large]

    @lru_cache(maxsize=None)
    def solve(
        index: int, used_mask: int,
    ) -> tuple[tuple[int, int, int], tuple[tuple[int, int], ...]]:
        if index == len(small):
            return (0, 0, 0), ()
        best = None
        for candidate in range(len(large)):
            if used_mask & (1 << candidate):
                continue
            tail_cost, tail_pairs = solve(
                index + 1, used_mask | (1 << candidate),
            )
            pair_cost = _cost(
                small_values[index], large_values[candidate], index, candidate,
            )
            value = (
                _add_cost(pair_cost, tail_cost),
                ((index, candidate), *tail_pairs),
            )
            if best is None or value < best:
                best = value
        if best is None:
            raise SourceCoreError("assignment DP found no candidate")
        return best

    total_cost, assignments = solve(0, 0)
    pairs: list[tuple[Row, Row]] = []
    for small_index, large_index in assignments:
        if swapped:
            pairs.append((large[large_index], small[small_index]))
        else:
            pairs.append((small[small_index], large[large_index]))
    pairs.sort(key=lambda pair: (pair[0].source_index, pair[1].source_index))
    left_only = [
        row for row in left if all(row is not pair[0] for pair in pairs)
    ]
    right_only = [
        row for row in right if all(row is not pair[1] for pair in pairs)
    ]
    return pairs, left_only, right_only, total_cost


def _pair(
    left: Sequence[Row],
    right: Sequence[Row],
    key: Callable[[Row], tuple[str, ...]],
    projector: Callable[[Row], tuple[str, ...]],
) -> tuple[
    list[tuple[Row, Row]], list[Row], list[Row], dict[str, object],
]:
    left_groups: dict[tuple[str, ...], list[Row]] = defaultdict(list)
    right_groups: dict[tuple[str, ...], list[Row]] = defaultdict(list)
    for row in left:
        left_groups[key(row)].append(row)
    for row in right:
        right_groups[key(row)].append(row)

    pairs: list[tuple[Row, Row]] = []
    left_only: list[Row] = []
    right_only: list[Row] = []
    assignment_cost = (0, 0, 0)
    duplicate_traces: list[dict[str, object]] = []
    for identity in sorted(set(left_groups) | set(right_groups)):
        group_left = left_groups.get(identity, [])
        group_right = right_groups.get(identity, [])
        group_pairs, group_left_only, group_right_only, group_cost = (
            _assign_group(group_left, group_right, projector)
        )
        pairs.extend(group_pairs)
        left_only.extend(group_left_only)
        right_only.extend(group_right_only)
        assignment_cost = _add_cost(assignment_cost, group_cost)
        if len(group_left) > 1 or len(group_right) > 1:
            duplicate_traces.append({
                "identity": identity,
                "left": [row.source_ref for row in group_left],
                "right": [row.source_ref for row in group_right],
                "pairs": [
                    [left_row.source_ref, right_row.source_ref]
                    for left_row, right_row in group_pairs
                ],
                "left_only": [row.source_ref for row in group_left_only],
                "right_only": [row.source_ref for row in group_right_only],
                "cost": group_cost,
            })
    pairs.sort(
        key=lambda pair: (
            pair[0].route, pair[0].source_index, pair[1].source_index,
        )
    )
    left_only.sort(key=lambda row: (row.route, row.source_index))
    right_only.sort(key=lambda row: (row.route, row.source_index))
    return pairs, left_only, right_only, {
        "left_key_groups": len(left_groups),
        "right_key_groups": len(right_groups),
        "duplicate_groups": len(duplicate_traces),
        "assignment_cost": assignment_cost,
        "duplicate_trace_sha256": _sha_bytes(_json_bytes(duplicate_traces)),
        "duplicate_traces": duplicate_traces,
    }


def _comparison(
    left: Sequence[Row],
    right: Sequence[Row],
    label: str,
    key_policy: str,
) -> tuple[dict[str, object], list[tuple[Row, Row]]]:
    if key_policy == "same_source_base_pm":
        key = lambda row: (row.route, row.county, row.pm_base)
        projector = _same_source_values
        left_values = right_values = _same_source_values
        fields = SAME_SOURCE_FIELDS
        asserted = set(fields)
    elif key_policy == "full_pm_vs_tsn":
        key = lambda row: (row.route, row.county, row.pm_full)
        projector = _vs_tsn_values
        left_values = right_values = _vs_tsn_values
        fields = VALUE_FIELDS
        asserted = set(ASSERTED_VS_TSN_FIELDS)
    else:
        raise SourceCoreError(f"unknown key policy: {key_policy}")

    pairs, left_only, right_only, pairing = _pair(
        left, right, key, projector,
    )
    all_field_counts: Counter[str] = Counter()
    asserted_field_counts: Counter[str] = Counter()
    all_differing_rows = 0
    asserted_differing_rows = 0
    differing_pairs: list[dict[str, object]] = []
    for left_row, right_row in pairs:
        left_projected = left_values(left_row)
        right_projected = right_values(right_row)
        differing = [
            field
            for field, a, b in zip(
                fields, left_projected, right_projected, strict=True,
            )
            if a != b
        ]
        asserted_differing = [
            field for field in differing if field in asserted
        ]
        if differing:
            all_differing_rows += 1
            all_field_counts.update(differing)
        if asserted_differing:
            asserted_differing_rows += 1
            asserted_field_counts.update(asserted_differing)
        if differing:
            differing_pairs.append({
                "identity": key(left_row),
                "left_ref": left_row.source_ref,
                "right_ref": right_row.source_ref,
                "left_pm_full": left_row.pm_full,
                "right_pm_full": right_row.pm_full,
                "left_values": left_projected,
                "right_values": right_projected,
                "differing_fields": differing,
                "asserted_differing_fields": asserted_differing,
                "left_record_kind": left_row.record_kind,
                "right_record_kind": right_row.record_kind,
                "left_source_role": left_row.source_role,
                "right_source_role": right_row.source_role,
            })
    result = {
        "label": label,
        "key_policy": key_policy,
        "fields": list(fields),
        "asserted_fields": sorted(asserted),
        "context_fields": [field for field in fields if field not in asserted],
        "left_rows": len(left),
        "right_rows": len(right),
        "paired_rows": len(pairs),
        "left_only_rows": len(left_only),
        "right_only_rows": len(right_only),
        "all_field_differing_rows": all_differing_rows,
        "all_field_difference_cells": sum(all_field_counts.values()),
        "all_field_difference_counts": dict(sorted(all_field_counts.items())),
        "asserted_differing_rows": asserted_differing_rows,
        "asserted_difference_cells": sum(asserted_field_counts.values()),
        "asserted_field_difference_counts": dict(
            sorted(asserted_field_counts.items())
        ),
        "left_only": [
            {
                "ref": row.source_ref, "route": row.route,
                "county": row.county, "pm": row.pm_full,
                "values": projector(row), "record_kind": row.record_kind,
                "source_role": row.source_role,
            }
            for row in left_only
        ],
        "right_only": [
            {
                "ref": row.source_ref, "route": row.route,
                "county": row.county, "pm": row.pm_full,
                "values": projector(row), "record_kind": row.record_kind,
                "source_role": row.source_role,
            }
            for row in right_only
        ],
        "differing_pairs": differing_pairs,
        "pairing": pairing,
    }
    return result, pairs


def _observed_contract(result: Mapping[str, object]) -> dict[str, object]:
    return {
        "shape": (
            result["left_rows"], result["right_rows"], result["paired_rows"],
            result["left_only_rows"], result["right_only_rows"],
        ),
        "all": (
            result["all_field_differing_rows"],
            result["all_field_difference_cells"],
            result["all_field_difference_counts"],
        ),
        "asserted": (
            result["asserted_differing_rows"],
            result["asserted_difference_cells"],
            result["asserted_field_difference_counts"],
        ),
    }


def _check_expected_leg(
    label: str,
    result: Mapping[str, object],
    expected: Mapping[str, object],
) -> None:
    observed = _observed_contract(result)
    for key in ("shape", "all", "asserted"):
        _require(
            observed[key] == expected[key],
            f"{label} {key} drift: {observed[key]} != {expected[key]}",
        )


def _unknown_raw_publication_ledger(rows: Sequence[Row]) -> dict[str, object]:
    records = [{
        "source_ref": row.source_ref,
        "source_index": row.source_index,
        "route": row.route,
        "county": row.county,
        "pm": row.pm_full,
        "values": row.values,
        "record_kind": row.record_kind,
        "source_role": row.source_role,
        "source_format": row.source_format,
        "provenance": row.provenance,
    } for row in rows]
    return {
        "rows": len(records),
        "all_blank_county": all(not row.county for row in rows),
        "all_equate_annotations": all(row.record_kind == "equate" for row in rows),
        "ordered_ledger_sha256": _sha_bytes(_json_bytes(records)),
        "records": records,
    }


def _complete_raw_publication_leg(
    keyable: Mapping[str, object], unknown_rows: Sequence[Row],
) -> dict[str, object]:
    ledger = _unknown_raw_publication_ledger(unknown_rows)
    _require(
        ledger["rows"] == 46
        and ledger["all_blank_county"] is True
        and ledger["all_equate_annotations"] is True
        and ledger["ordered_ledger_sha256"]
        == EXPECTED_UNKNOWN_RAW_PUBLICATION_SHA256,
        "complete raw-publication unknown-County ledger drift",
    )
    unknown_one_sided = [{
        "ref": row.source_ref,
        "route": row.route,
        "county": row.county,
        "pm": row.pm_full,
        "values": _vs_tsn_values(row),
        "record_kind": row.record_kind,
        "source_role": row.source_role,
        "unkeyed_reason": "authoritative raw EQUATES annotation precedes County context",
        "provenance": row.provenance,
    } for row in unknown_rows]
    right_only = sorted(
        [*keyable["right_only"], *unknown_one_sided],
        key=lambda item: (
            item["route"], item["ref"], item["pm"], item["record_kind"],
        ),
    )
    complete = dict(keyable)
    complete.update({
        "label": keyable["label"] + " (complete 69,804-row raw publication)",
        "key_policy": (
            "keyable rows: route + normalized County + complete printed PM; "
            "46 blank-County raw annotations are explicitly unkeyed TSN-only rows"
        ),
        "right_rows": keyable["right_rows"] + len(unknown_rows),
        "right_only_rows": keyable["right_only_rows"] + len(unknown_rows),
        "right_only": right_only,
        "keyable_semantic_contract": {
            "right_rows": keyable["right_rows"],
            "paired_rows": keyable["paired_rows"],
            "left_only_rows": keyable["left_only_rows"],
            "right_only_rows": keyable["right_only_rows"],
            "meaning": (
                "69,758 raw records with printed County can participate in the "
                "semantic key; this is not the complete raw publication shape."
            ),
        },
        "unkeyed_raw_tsn_only": ledger,
        "complete_raw_publication": True,
    })
    return complete


def _raw_vs_normalized(
    raw: Sequence[Row], normalized: Sequence[Row],
) -> dict[str, object]:
    _require(len(raw) == len(normalized), "raw/normalized row-count drift")
    fields = ("Route", "County", "PM", *VALUE_FIELDS)
    differences: list[dict[str, object]] = []
    counts: Counter[str] = Counter()
    for ordinal, (left, right) in enumerate(zip(raw, normalized, strict=True), 1):
        left_values = (left.route, left.county, left.pm_full, *left.values)
        right_values = (right.route, right.county, right.pm_full, *right.values)
        differing = [
            field
            for field, a, b in zip(fields, left_values, right_values, strict=True)
            if a != b
        ]
        if differing:
            counts.update(differing)
            differences.append({
                "ordinal": ordinal,
                "raw_ref": left.source_ref,
                "normalized_ref": right.source_ref,
                "raw_values": left_values,
                "normalized_values": right_values,
                "differing_fields": differing,
            })
    return {
        "rows": len(raw),
        "differing_rows": len(differences),
        "difference_cells": sum(counts.values()),
        "field_difference_counts": dict(sorted(counts.items())),
        "differences": differences,
    }


def _same_source_events(
    excel: Sequence[Row], pdf: Sequence[Row],
) -> dict[str, object]:
    excel_groups: dict[tuple[str, str, str], list[Row]] = defaultdict(list)
    pdf_groups: dict[tuple[str, str, str], list[Row]] = defaultdict(list)
    for row in excel:
        excel_groups[(row.route, row.county, row.pm_base)].append(row)
    for row in pdf:
        pdf_groups[(row.route, row.county, row.pm_base)].append(row)
    excel_by_occurrence = {
        (*identity, index): row
        for identity, rows in excel_groups.items()
        for index, row in enumerate(rows, 1)
    }
    pdf_by_route: dict[str, list[Row]] = defaultdict(list)
    for row in pdf:
        pdf_by_route[row.route].append(row)
    pdf_identity = {
        id(row): (*identity, index)
        for identity, rows in pdf_groups.items()
        for index, row in enumerate(rows, 1)
    }

    classes: Counter[str] = Counter()
    events: list[dict[str, object]] = []
    linked_pdf_e_refs: set[str] = set()
    for route, rows in pdf_by_route.items():
        for index, annotation in enumerate(rows):
            if not annotation.values[-1].startswith("EQUATES TO"):
                continue
            _require(index + 1 < len(rows), f"final-row equate: {annotation.source_ref}")
            following = rows[index + 1]
            annotation_identity = pdf_identity[id(annotation)]
            following_identity = pdf_identity[id(following)]
            excel_annotation = excel_by_occurrence[annotation_identity]
            excel_following = excel_by_occurrence[following_identity]
            _require(
                not annotation.pm_suffix and following.pm_suffix == "E",
                f"PDF equate topology drift: {annotation.source_ref}",
            )
            linked_pdf_e_refs.add(following.source_ref)
            state = (
                excel_annotation.pm_suffix, annotation.pm_suffix,
                excel_following.pm_suffix, following.pm_suffix,
            )
            categories = {
                ("", "", "E", "E"): "same_convention",
                ("E", "", "", "E"):
                    "suffix_moved_between_annotation_and_following",
                ("", "", "", "E"): "pdf_only_suffix_on_following",
            }
            _require(state in categories, f"unknown equate state {state}")
            category = categories[state]
            classes[category] += 1
            events.append({
                "category": category,
                "pdf_annotation_ref": annotation.source_ref,
                "excel_annotation_ref": excel_annotation.source_ref,
                "pdf_following_ref": following.source_ref,
                "excel_following_ref": excel_following.source_ref,
                "annotation_identity": annotation_identity,
                "following_identity": following_identity,
                "excel_annotation_suffix": excel_annotation.pm_suffix,
                "pdf_annotation_suffix": annotation.pm_suffix,
                "excel_following_suffix": excel_following.pm_suffix,
                "pdf_following_suffix": following.pm_suffix,
            })
    unlinked_pdf_e = [
        row.source_ref for row in pdf
        if row.pm_suffix == "E" and row.source_ref not in linked_pdf_e_refs
    ]
    route152 = [
        {
            "occurrence": index,
            "excel_ref": excel_row.source_ref,
            "pdf_ref": pdf_row.source_ref,
            "excel_suffix": excel_row.pm_suffix,
            "pdf_suffix": pdf_row.pm_suffix,
            "excel_values": excel_row.values,
            "pdf_values": pdf_row.values,
        }
        for index, (excel_row, pdf_row) in enumerate(zip(
            excel_groups[("152", "SCR", "T003.273")],
            pdf_groups[("152", "SCR", "T003.273")],
            strict=True,
        ), 1)
    ]
    return {
        "equate_events": len(events),
        "classes": dict(sorted(classes.items())),
        "unlinked_pdf_e_rows": unlinked_pdf_e,
        "event_ledger_sha256": _sha_bytes(_json_bytes(events)),
        "route_152_scr_t003_273_occurrences": route152,
        "events": events,
    }


def _product_description(value: object) -> str:
    """Model the exact bound product rewrite without importing product code."""
    text = _text(value)
    text = re.sub(r"[\t\n\r\f\v]", " ", text).strip()
    return re.sub(r"\s+", " ", ROUTE_PREFIX_RE.sub("", text)).strip()


def _prefix_population(rows: Sequence[Row], label: str) -> dict[str, object]:
    selected = [row for row in rows if ROUTE_PREFIX_RE.match(row.values[-1])]
    records: list[dict[str, object]] = []
    owning = cross_route = changed = 0
    nested: list[dict[str, object]] = []
    for row in selected:
        match = ROUTE_PREFIX_RE.match(row.values[-1])
        if match is None:
            raise SourceCoreError("prefix population internal error")
        token_route = _route(match.group(1))
        relation = "owning-route" if token_route == row.route else "cross-route"
        owning += relation == "owning-route"
        cross_route += relation == "cross-route"
        product = _product_description(row.values[-1])
        changed += product != row.values[-1]
        nested_prefix = ROUTE_PREFIX_RE.match(product) is not None
        record = {
            "identity": list(row.identity),
            "description": row.values[-1],
            "prefix_token": match.group(1),
            "relation": relation,
            "product_description": product,
            "nested_prefix_survives": nested_prefix,
        }
        records.append(record)
        if nested_prefix:
            nested.append({**record, "source_ref": row.source_ref})
    context_multiset = sorted([
        list(row.identity) + [row.values[-1]] for row in selected
    ])
    description_multiset = sorted(row.values[-1] for row in selected)
    return {
        "label": label,
        "rows": len(selected),
        "distinct_descriptions": len(set(description_multiset)),
        "owning_route_prefixes": owning,
        "cross_route_prefixes": cross_route,
        "changed_by_product": changed,
        "numeric_prefixes_remaining_after_product": len(nested),
        "context_multiset_sha256": _sha_bytes(
            _json_bytes(context_multiset, newline=False)
        ),
        "description_multiset_sha256": _sha_bytes(
            _json_bytes(description_multiset, newline=False)
        ),
        "context_multiset": context_multiset,
        "description_multiset": description_multiset,
        "nested_prefix_records": nested,
        "records": sorted(
            records,
            key=lambda item: (
                tuple(item["identity"]), item["description"], item["relation"],
            ),
        ),
    }


def _collapsed_duplicate_distinctions(rows: Sequence[Row]) -> list[dict[str, object]]:
    groups: dict[tuple[str, str, str], list[Row]] = defaultdict(list)
    for row in rows:
        groups[row.identity].append(row)
    collapsed: list[dict[str, object]] = []
    for identity, group in groups.items():
        product_groups: dict[str, set[str]] = defaultdict(set)
        for row in group:
            product_groups[_product_description(row.values[-1])].add(row.values[-1])
        for product, originals in product_groups.items():
            if len(originals) > 1 and any(
                ROUTE_PREFIX_RE.match(item) for item in originals
            ):
                collapsed.append({
                    "identity": list(identity),
                    "product_description": product,
                    "source_descriptions": sorted(originals),
                })
    return sorted(collapsed, key=lambda item: tuple(item["identity"]))


def _padding_artifacts(rows: Sequence[Row]) -> list[dict[str, object]]:
    artifacts: list[dict[str, object]] = []
    for row in rows:
        _require(row.source_role == "tsmis", "padding census received TSN row")
        raw = _space(
            row.raw_values[8], xlsx=row.source_format == "xlsx",
        )
        match = PREFIX_WITH_PADDING_RE.match(raw)
        if match is None or _route(match.group(1)) != row.route:
            continue
        prefix = ROUTE_PREFIX_RE.match(raw)
        if prefix is None:
            raise SourceCoreError("padding match lost its owning route prefix")
        artifacts.append({
            "identity": list(row.identity),
            "source_ref": row.source_ref,
            "raw_description": row.raw_values[8],
            "untrimmed_projection": raw[prefix.end():],
            "corrected_projection": row.values[-1][match.end():].lstrip()
            if ROUTE_PREFIX_RE.match(row.values[-1]) else _semantic_description(
                row.raw_values[8], row.route,
                xlsx=row.source_format == "xlsx",
                strip_tsmis_route_label=True,
            ),
            "separator_whitespace_length": len(match.group(2)),
        })
    return sorted(artifacts, key=lambda item: tuple(item["identity"]))


def _false_clean_ledger(
    pairs: Sequence[tuple[Row, Row]],
) -> dict[str, object]:
    records: list[dict[str, object]] = []
    references: list[dict[str, object]] = []
    for left, right in pairs:
        left_description = _vs_tsn_values(left)[-1]
        right_description = _vs_tsn_values(right)[-1]
        if (
            left_description == right_description
            or ROUTE_PREFIX_RE.match(right_description) is None
            or _product_description(right_description) != left_description
        ):
            continue
        _require(
            _product_description(left.raw_values[8])
            == _product_description(right_description),
            f"modeled product sides do not false-clean at {left.identity}",
        )
        records.append({
            "identity": list(left.identity),
            "tsmis": left_description,
            "tsn": right_description,
        })
        references.append({
            "identity": list(left.identity),
            "tsmis_ref": left.source_ref,
            "tsn_ref": right.source_ref,
        })
    content_sorted = sorted(
        records,
        key=lambda item: (
            tuple(item["identity"]), item["tsmis"], item["tsn"],
        ),
    )
    return {
        "rows": len(records),
        "wire_bytes": len(_json_bytes(records, newline=False)),
        "sha256": _sha_bytes(_json_bytes(records, newline=False)),
        "content_sorted_sha256": _sha_bytes(
            _json_bytes(content_sorted, newline=False)
        ),
        "records": records,
        "references": references,
    }


def _parse_raw_tsn() -> tuple[
    list[dict[str, object]], list[dict[str, object]],
    list[list[dict[str, object]]], dict[str, object],
]:
    all_records: list[dict[str, object]] = []
    documents: list[dict[str, object]] = []
    document_records: list[list[dict[str, object]]] = []
    captured_identities: list[dict[str, object]] = []
    raw_root = PRIVATE_ROOT / "authoritative_tsn_pdf"
    for name, size, digest in stage6.RAW_BINDINGS:
        path = raw_root / name
        payload = path.read_bytes()
        _require(
            len(payload) == size and _sha_bytes(payload) == digest,
            f"private authoritative TSN identity drift: {name}",
        )
        captured_identities.append({
            "name": name, "bytes": len(payload), "sha256": _sha_bytes(payload),
        })
        records, document = stage6._parse_document(name, payload)
        all_records.extend(records)
        document_records.append(records)
        documents.append(document)
    return (
        stage6._sorted_source(all_records), documents, document_records,
        {
            "members": len(captured_identities),
            "bytes": sum(item["bytes"] for item in captured_identities),
            "ordered_identity_sha256": _sha_bytes(
                _json_bytes(captured_identities)
            ),
            "identities": captured_identities,
            "parse_uses_this_exact_payload_capture": True,
        },
    )


def _read_normalized_tsn():
    spec = xlsx_stream.SheetSpec(
        stage6.SHEET_NAME,
        tuple(xlsx_stream.ColumnSpec(header) for header in stage6.HEADERS),
        exact_schema=True,
    )
    return xlsx_stream.read_sheet(
        NORMALIZED_TSN,
        spec,
        limits=xlsx_stream.XlsxLimits(max_source_bytes=32 * 1024 * 1024),
    )


def _raw_event_topology(
    document_records: Sequence[Sequence[Mapping[str, object]]],
) -> dict[str, object]:
    events: list[dict[str, object]] = []
    for records in document_records:
        for index, annotation in enumerate(records):
            if annotation["kind"] != "equate":
                continue
            _require(index + 1 < len(records), "raw TSN equate is final document row")
            following = records[index + 1]
            _require(
                following["kind"] == "data"
                and str(following["pm"]).endswith("E")
                and following["member"] == annotation["member"]
                and following["district"] == annotation["district"]
                and following["route"] == annotation["route"]
                and following["direction"] == annotation["direction"],
                f"raw TSN equate topology drift: {annotation}",
            )
            events.append({
                "annotation": {
                    key: annotation[key] for key in (
                        "member", "district", "route", "direction", "county",
                        "pm", "physical_page", "printed_page", "line", "top",
                    )
                },
                "following": {
                    key: following[key] for key in (
                        "member", "district", "route", "direction", "county",
                        "pm", "physical_page", "printed_page", "line", "top",
                    )
                },
            })
    return {
        "equate_events": len(events),
        "all_immediately_followed_by_owned_e_row": len(events) == 998,
        "ledger_sha256": _sha_bytes(_json_bytes(events)),
        "events": events,
    }


def _fresh_stage6_proof(
    records: Sequence[Mapping[str, object]],
    documents: Sequence[Mapping[str, object]],
    normalized_sheet,
    accepted: Mapping[str, object],
) -> dict[str, object]:
    source_rows = [stage6._source_row(record) for record in records]
    projected_records = [record for record in records if record["county"] is not None]
    projected_rows = [stage6._project_record(record) for record in projected_records]
    _require(not any(row is None for row in projected_rows), "projection omitted known County")
    projection = stage6._compare_projection(projected_rows, normalized_sheet.rows)

    source_provenance_rows = [(
        record["member"], record["physical_page"], record["printed_page"],
        record["line"], record["top"], record["raw_text"],
    ) for record in records]
    source_provenance_headers = (
        "Member", "Physical Page", "Printed Page", "Line", "Top", "Raw Text",
    )
    metadata_rows = []
    for document in documents:
        times = sorted({
            claim["generation_time"] for claim in document["data_page_claims"]
        })
        metadata_rows.append((
            document["member"], document["district"], "OTM22025",
            "Highway Locations", "15-SEP-25", "15 SEP 2025",
            "|".join(times), document["policy_sha256"],
            json.dumps(
                document["pdf_metadata"], sort_keys=True, separators=(",", ":"),
            ),
        ))
    metadata_headers = (
        "Member", "District", "Report ID", "Report Title",
        "Cover Reference Date", "Data Reference Date", "Generation Time",
        "Policy SHA256", "PDF Metadata",
    )
    fresh_digests = {
        "raw_source_digests": stage6._dataset_digests(
            stage6.SOURCE_HEADERS, source_rows,
        ),
        "raw_source_provenance_digests": stage6._dataset_digests(
            source_provenance_headers, source_provenance_rows,
        ),
        "document_metadata_digests": stage6._dataset_digests(
            metadata_headers, metadata_rows,
        ),
        "independently_projected_digests": stage6._dataset_digests(
            stage6.HEADERS, projected_rows,
        ),
        "normalized_digests": stage6._dataset_digests(
            stage6.HEADERS, [row.values for row in normalized_sheet.rows],
        ),
    }
    for label, value in fresh_digests.items():
        _require(value == accepted[label], f"fresh Stage-6 {label} drift")
    _require(
        projection == accepted["projection_comparison"],
        "fresh raw-to-normalized typed projection drifted from accepted Stage 6",
    )
    _require(
        stage6.FIELD_DISPOSITIONS == accepted["field_dispositions"]
        and stage6._field_coverage() == accepted["field_coverage"],
        "accepted provenance-role disposition drift",
    )

    pointer_mismatches = [
        item for item in projection["mismatches"]
        if item["field"] == "Distance To Next Point"
        and item["expected"] in (["str", "*P*"], ["str", "-------->"])
        and item["actual"] == ["null"]
    ]
    description_mismatches = [
        item for item in projection["mismatches"]
        if item["field"] == "Description"
    ]
    pointer_records = [
        record for record in projected_records
        if record["distance"] in ("*P*", "-------->")
    ]
    unknown_equates = [
        record for record in records
        if record["kind"] == "equate" and record["county"] is None
    ]
    continuations = [
        item for document in documents for item in document["continuations"]
    ]
    punctuation_exact = (
        len(description_mismatches) == 1
        and projected_records[description_mismatches[0]["ordinal"]]["district"] == "09"
        and projected_records[description_mismatches[0]["ordinal"]]["county"] == "KER"
        and projected_records[description_mismatches[0]["ordinal"]]["route"] == "014"
        and projected_records[description_mismatches[0]["ordinal"]]["pm"] == "018.365"
        and description_mismatches[0]["expected"] == [
            "str", "KEMWATER CHEMICAL PLANT - RT/FRONTAGE ROAD - LT.",
        ]
        and description_mismatches[0]["actual"] == [
            "str", "KEMWATER CHEMICAL PLANT - RT/FRONTAGE, ROAD - LT.",
        ]
    )
    classified = {
        "pre_county_equates_dropped": {
            "count": len(unknown_equates),
            "manifest_sha256": stage6._sha(stage6._json_bytes([{
                key: record[key] for key in (
                    "member", "district", "route", "direction", "physical_page",
                    "printed_page", "line", "pm", "raw_text",
                )
            } for record in unknown_equates])),
            "manifest": [{
                key: record[key] for key in (
                    "member", "district", "route", "direction", "physical_page",
                    "printed_page", "line", "pm", "raw_text",
                )
            } for record in unknown_equates],
            "finding": "CMP-AUD-158",
        },
        "distance_pointer_tokens_blanked": {
            "count": len(pointer_mismatches),
            "domain": dict(sorted(Counter(
                record["distance"] for record in pointer_records
            ).items())),
            "manifest_sha256": stage6._sha(stage6._json_bytes([{
                key: record[key] for key in (
                    "member", "district", "route", "county", "pm",
                    "physical_page", "printed_page", "line", "distance",
                )
            } for record in pointer_records])),
            "manifest": [{
                key: record[key] for key in (
                    "member", "district", "route", "county", "pm",
                    "physical_page", "printed_page", "line", "distance",
                )
            } for record in pointer_records],
            "finding": "CMP-AUD-156",
        },
        "wrapped_description_invented_comma": {
            "count": len(description_mismatches),
            "exact": punctuation_exact,
            "mismatches": description_mismatches,
            "continuation_claims": continuations,
            "finding": "CMP-AUD-159",
        },
    }
    _require(
        classified == accepted["classified_projection_residue"],
        "fresh classified provenance-loss ledger drifted from accepted Stage 6",
    )
    _require(
        len(records) == 69_804
        and sum(record["kind"] == "data" for record in records) == 68_806
        and sum(record["kind"] == "equate" for record in records) == 998
        and len(unknown_equates) == 46
        and len(pointer_mismatches) == 565
        and punctuation_exact
        and len(normalized_sheet.rows) == 69_758,
        "fresh Stage-6 raw/normalized census drift",
    )
    return {
        "fresh_digests": fresh_digests,
        "projection_comparison": projection,
        "classified_projection_residue": classified,
        "field_dispositions": stage6.FIELD_DISPOSITIONS,
        "field_coverage": stage6._field_coverage(),
        "source_only_or_provenance_roles": sorted(
            field for field, disposition in stage6.FIELD_DISPOSITIONS.items()
            if disposition["kind"] in {
                "source_only", "source_only_metadata", "relational",
                "audit_provenance",
            }
        ),
        "all_fresh_ledgers_equal_accepted_stage6": True,
    }


def _diagnostic_summary(
    rows: Sequence[tsmis_source.SourceRow], diagnostic: Mapping[str, object],
) -> dict[str, object]:
    summary = {
        "rows": len(rows),
        "members": diagnostic["members"],
        "ordered_raw_rows_sha256": tsmis_source._rows_digest(rows),
        "diagnostic_sha256": _sha_bytes(_json_bytes(diagnostic)),
    }
    for key in (
        "pages", "data_pages", "description_fragments", "members_with_trailer",
        "all_cells_string_or_null", "header_anchor_x0_ranges",
    ):
        if key in diagnostic:
            summary[key] = diagnostic[key]
    return summary


def _source_row_wire(row: tsmis_source.SourceRow) -> tuple[object, ...]:
    return (
        row.member, row.route, row.source_index, row.source_ref,
        row.page, row.top, row.values,
    )


def _edition_proof(
    historical_excel: Sequence[tsmis_source.SourceRow],
    current_excel: Sequence[tsmis_source.SourceRow],
    historical_pdf: Sequence[tsmis_source.SourceRow],
    current_pdf: Sequence[tsmis_source.SourceRow],
) -> dict[str, object]:
    excel_delta = tsmis_source._parity(
        historical_excel, current_excel,
        "historical July-8 Excel vs current July-9 Excel",
    )
    pdf_delta = tsmis_source._parity(
        historical_pdf, current_pdf,
        "historical PDF capture vs current PDF capture",
    )
    expected_changed = {
        ("002", "LA", "", "014.348", "occurrence:1"):
            "002/EB ON FR GLENDALE BLVD",
        ("010", "LA", "", "014.820", "occurrence:1"):
            "010/SEG EB CONN OFF TO GRAND/18TH",
        ("037", "SON", "", "003.809", "occurrence:1"):
            "037/WB ON FR SB RTE 121",
        ("037", "SON", "", "003.981", "occurrence:1"):
            "037/WB OFF TO NB RTE 121",
        ("101", "SBT", "", "002.999", "occurrence:1"):
            "101/SB TO 156 TWO WAY CONN",
    }
    observed_changed = {
        tuple(item["identity"]): _space(item["right_values"][8], xlsx=True)
        for item in excel_delta["differing_pairs"]
    }
    _require(observed_changed == expected_changed, "five paired edition updates drift")
    _require(
        excel_delta["paired_rows"] == 60_493
        and excel_delta["left_only_rows"] == 0
        and excel_delta["right_only_rows"] == 1
        and excel_delta["display_differing_rows"] == 5
        and excel_delta["display_field_difference_counts"] == {"Description": 5},
        "current-vs-historical Excel shape drift",
    )
    new_row = excel_delta["right_only"][0]
    _require(
        tuple(new_row["identity"])
        == ("010", "LA", "", "014.814", "occurrence:1")
        and _space(new_row["values"][8], xlsx=True) == "010/EB ON FR VERMONT",
        "sixth edition update (new route-010 row) drift",
    )
    _require(
        pdf_delta["paired_rows"] == 60_493
        and pdf_delta["left_only_rows"] == 0
        and pdf_delta["right_only_rows"] == 0
        and pdf_delta["raw_differing_rows"] == 0,
        "byte-identical historical/current PDF parse drift",
    )
    _require(
        [_source_row_wire(row) for row in historical_pdf]
        == [_source_row_wire(row) for row in current_pdf],
        "direct historical/current PDF row reconstructions differ",
    )
    return {
        "paired_description_updates": excel_delta["differing_pairs"],
        "new_current_row": new_row,
        "source_updates": 6,
        "historical_current_pdf_rows_exact": True,
        "historical_current_pdf_parity": pdf_delta,
        "excel_parity": excel_delta,
    }


def _same_source_omissions(result: Mapping[str, object]) -> dict[str, object]:
    paired = []
    for item in result["differing_pairs"]:
        if (
            "Description" in item["differing_fields"]
            and item["left_values"][-1] == ""
            and item["right_values"][-1]
        ):
            paired.append(item)
    expected = {
        ("002", "LA", "014.348"),
        ("010", "LA", "014.820"),
        ("037", "SON", "003.981"),
        ("101", "SBT", "002.999"),
    }
    observed = {
        (item["identity"][0], item["identity"][1], item["identity"][2])
        for item in paired
    }
    _require(observed == expected, "four paired PDF Description omissions drift")
    _require(
        len(result["right_only"]) == 1
        and result["right_only"][0]["route"] == "010"
        and result["right_only"][0]["county"] == "LA"
        and result["right_only"][0]["pm"] == "014.814",
        "Excel-only described row drift",
    )
    return {
        "paired_pdf_description_omissions": paired,
        "excel_only_described_row": result["right_only"][0],
        "independent_unrepresented_claims": 5,
    }


def _validate_stage6_acceptance(
    result: Mapping[str, object], decision: Mapping[str, object],
) -> None:
    _require(
        result.get("stage6_family_audit_complete") is True
        and result.get("unexplained_projection_residue_count") == 0,
        "accepted Stage-6 result lost its terminal family contract",
    )
    tracked = decision.get("tracked_identities", {})
    _require(
        decision.get("decision") == "accepted_stage6_family_audit"
        and tracked.get("result", {}).get("sha256")
        == STATIC_FILE_BINDINGS["accepted_stage6_result"]["sha256"]
        and tracked.get("normalized", {}).get("sha256")
        == STATIC_FILE_BINDINGS["accepted_normalized_tsn"]["sha256"]
        and tracked.get("oracle", {}).get("sha256")
        == STATIC_FILE_BINDINGS["stage6_oracle"]["sha256"]
        and tracked.get("reader", {}).get("sha256")
        == STATIC_FILE_BINDINGS["xlsx_reader"]["sha256"],
        "accepted Stage-6 detached decision chain drift",
    )


def _mutation_rejection(
    label: str,
    expected: object,
    observe: Callable[[], object],
    *,
    mutated_input: object,
) -> dict[str, object]:
    expected_digest = _sha_bytes(_json_bytes(expected))
    mutated_input_digest = _sha_bytes(_json_bytes(mutated_input))
    try:
        observed = observe()
    except SourceCoreError as error:
        return {
            "rejected": True,
            "failure_reason": str(error),
            "expected_contract_sha256": expected_digest,
            "observed_contract_sha256": None,
            "mutated_input_sha256": mutated_input_digest,
        }
    observed_digest = _sha_bytes(_json_bytes(observed))
    try:
        _require(observed == expected, f"{label}: mutated contract mismatch")
    except SourceCoreError as error:
        return {
            "rejected": True,
            "failure_reason": str(error),
            "expected_contract_sha256": expected_digest,
            "observed_contract_sha256": observed_digest,
            "mutated_input_sha256": mutated_input_digest,
        }
    raise SourceCoreError(f"negative mutation escaped the {label} gate")


def _pointer_contract(rows: Sequence[Row]) -> dict[str, object]:
    records = [{
        "source_ref": row.source_ref,
        "route": row.route,
        "county": row.county,
        "pm": row.pm_full,
        "token": row.values[3],
    } for row in rows if row.values[3] in ("*P*", "-------->")]
    return {
        "rows": len(records),
        "domain": dict(sorted(Counter(item["token"] for item in records).items())),
        "ordered_ledger_sha256": _sha_bytes(_json_bytes(records)),
    }


def _prefix_contract(result: Mapping[str, object]) -> dict[str, object]:
    return {
        "rows": result["rows"],
        "owning_route_prefixes": result["owning_route_prefixes"],
        "cross_route_prefixes": result["cross_route_prefixes"],
        "changed_by_product": result["changed_by_product"],
        "numeric_prefixes_remaining_after_product": (
            result["numeric_prefixes_remaining_after_product"]
        ),
        "context_multiset_sha256": result["context_multiset_sha256"],
        "description_multiset_sha256": result["description_multiset_sha256"],
    }


def _padding_contract(
    artifacts: Sequence[Mapping[str, object]],
) -> dict[str, object]:
    return {
        "rows": len(artifacts),
        "identities": [item["identity"] for item in artifacts],
        "untrimmed_sha256": _sha_bytes(_json_bytes([
            item["untrimmed_projection"] for item in artifacts
        ])),
        "corrected_sha256": _sha_bytes(_json_bytes([
            item["corrected_projection"] for item in artifacts
        ])),
    }


def _event_contract(events: Mapping[str, object]) -> dict[str, object]:
    return {
        "equate_events": events["equate_events"],
        "classes": events["classes"],
        "unlinked_pdf_e_rows": events["unlinked_pdf_e_rows"],
        "event_ledger_sha256": events["event_ledger_sha256"],
        "route152_suffixes": [
            [item["excel_suffix"], item["pdf_suffix"]]
            for item in events["route_152_scr_t003_273_occurrences"]
        ],
    }


def _edition_row_contract(
    historical: Sequence[Row], current: Sequence[Row],
) -> dict[str, object]:
    def index(rows: Sequence[Row]) -> dict[tuple[str, str, str, int], Row]:
        counts: Counter[tuple[str, str, str]] = Counter()
        result = {}
        for row in rows:
            base = row.route, row.county, row.pm_base
            counts[base] += 1
            result[(*base, counts[base])] = row
        return result
    old = index(historical)
    new = index(current)
    changed = [{
        "identity": identity,
        "historical_description": old[identity].values[-1],
        "current_description": new[identity].values[-1],
    } for identity in sorted(set(old) & set(new))
        if old[identity].values[-1] != new[identity].values[-1]]
    right_only = [{
        "identity": identity,
        "description": new[identity].values[-1],
    } for identity in sorted(set(new) - set(old))]
    return {
        "paired_description_updates": changed,
        "current_only_rows": right_only,
        "source_updates": len(changed) + len(right_only),
        "ledger_sha256": _sha_bytes(_json_bytes([changed, right_only])),
    }


def _dataset_role_contract(rows: Sequence[Row]) -> dict[str, object]:
    records = [{
        "dataset": row.dataset,
        "source_role": row.source_role,
        "source_format": row.source_format,
        "source_ref": row.source_ref,
    } for row in rows]
    return {
        "rows": len(records),
        "datasets": dict(sorted(Counter(item["dataset"] for item in records).items())),
        "source_roles": dict(sorted(Counter(
            item["source_role"] for item in records
        ).items())),
        "source_formats": dict(sorted(Counter(
            item["source_format"] for item in records
        ).items())),
        "ordered_ledger_sha256": _sha_bytes(_json_bytes(records)),
    }


def _negative_mutations(
    excel: Sequence[Row],
    pdf: Sequence[Row],
    historical_pdf: Sequence[Row],
    historical_excel: Sequence[Row],
    raw: Sequence[Row],
    raw_unknown: Sequence[Row],
    document_records: Sequence[Sequence[Mapping[str, object]]],
    source_events: Mapping[str, object],
    prefix: Mapping[str, object],
    padding: Mapping[str, Sequence[Mapping[str, object]]],
) -> dict[str, object]:
    probes: dict[str, dict[str, object]] = {}

    expected_events = _event_contract(source_events)
    route152_index = next(
        index for index, row in enumerate(excel)
        if row.route == "152" and row.county == "SCR"
        and row.pm_base == "T003.273" and row.pm_suffix == "E"
    )
    mutated_excel = list(excel)
    mutated_excel[route152_index] = replace(
        mutated_excel[route152_index], pm_suffix="",
    )
    probes["route152_suffix_event_gate"] = _mutation_rejection(
        "route-152 same-source event",
        expected_events,
        lambda: _event_contract(_same_source_events(mutated_excel, pdf)),
        mutated_input={
            "source_ref": mutated_excel[route152_index].source_ref,
            "pm_suffix": mutated_excel[route152_index].pm_suffix,
        },
    )

    historical_source = next(
        row for row in historical_excel
        if ROUTE_PREFIX_RE.match(_space(row.raw_values[8], xlsx=True))
        and _route(ROUTE_PREFIX_RE.match(
            _space(row.raw_values[8], xlsx=True)
        ).group(1)) == row.route
    )
    wrong_role = replace(historical_source, source_role="tsn")
    expected_projection = _vs_tsn_values(historical_source)
    probes["typed_source_role_gate"] = _mutation_rejection(
        "typed TSMIS source-role projection",
        expected_projection,
        lambda: _vs_tsn_values(wrong_role),
        mutated_input={
            "source_ref": wrong_role.source_ref,
            "source_role": wrong_role.source_role,
            "source_format": wrong_role.source_format,
        },
    )

    expected_historical_pdf_role = _dataset_role_contract(historical_pdf)
    mutated_historical_pdf = list(historical_pdf)
    mutated_historical_pdf[0] = replace(
        mutated_historical_pdf[0], dataset="current_tsmis_pdf",
    )
    probes["historical_pdf_dataset_role_gate"] = _mutation_rejection(
        "historical PDF typed dataset provenance",
        expected_historical_pdf_role,
        lambda: _dataset_role_contract(mutated_historical_pdf),
        mutated_input={
            "source_ref": historical_pdf[0].source_ref,
            "before": historical_pdf[0].dataset,
            "after": "current_tsmis_pdf",
        },
    )

    expected_pointer = _pointer_contract(raw)
    _require(
        expected_pointer["rows"] == 565
        and expected_pointer["domain"] == {"*P*": 283, "-------->": 282},
        "baseline pointer contract drift",
    )
    pointer_index = next(
        index for index, row in enumerate(raw)
        if row.values[3] in ("*P*", "-------->")
    )
    mutated_raw_pointer = list(raw)
    pointer_row = mutated_raw_pointer[pointer_index]
    mutated_raw_pointer[pointer_index] = replace(
        pointer_row, values=(*pointer_row.values[:3], "", pointer_row.values[4]),
    )
    probes["pointer_residue_ledger_gate"] = _mutation_rejection(
        "raw pointer residue ledger",
        expected_pointer,
        lambda: _pointer_contract(mutated_raw_pointer),
        mutated_input={
            "source_ref": pointer_row.source_ref,
            "before": pointer_row.values[3], "after": "",
        },
    )

    expected_unknown = _unknown_raw_publication_ledger(raw_unknown)
    mutated_unknown = list(raw_unknown[:-1])
    probes["unknown_county_manifest_and_census_gate"] = _mutation_rejection(
        "46-row blank-County raw-publication ledger",
        expected_unknown,
        lambda: _unknown_raw_publication_ledger(mutated_unknown),
        mutated_input={
            "deleted_source_ref": raw_unknown[-1].source_ref,
            "remaining_rows": len(mutated_unknown),
        },
    )

    expected_prefix = _prefix_contract(prefix)
    prefixed_index = next(
        index for index, row in enumerate(raw)
        if ROUTE_PREFIX_RE.match(row.values[-1])
    )
    prefixed_row = raw[prefixed_index]
    mutated_prefix_rows = list(raw)
    mutated_prefix_rows[prefixed_index] = replace(
        prefixed_row,
        values=(
            *prefixed_row.values[:-1],
            _product_description(prefixed_row.values[-1]),
        ),
    )
    probes["numeric_prefix_population_gate"] = _mutation_rejection(
        "154-row authoritative numeric-prefix population",
        expected_prefix,
        lambda: _prefix_contract(_prefix_population(
            mutated_prefix_rows, "mutated raw TSN",
        )),
        mutated_input={
            "source_ref": prefixed_row.source_ref,
            "before": prefixed_row.values[-1],
            "after": mutated_prefix_rows[prefixed_index].values[-1],
        },
    )

    expected_padding = _padding_contract(padding["current_tsmis_excel"])
    padding_identity = tuple(padding["current_tsmis_excel"][0]["identity"])
    padding_index = next(
        index for index, row in enumerate(excel)
        if row.identity == padding_identity
    )
    padding_row = excel[padding_index]
    mutated_raw_values = list(padding_row.raw_values)
    mutated_raw_values[8] = _space(mutated_raw_values[8], xlsx=True).replace(
        f"{padding_row.route}/ ", f"{padding_row.route}/", 1,
    )
    mutated_padding_rows = list(excel)
    mutated_padding_rows[padding_index] = replace(
        padding_row, raw_values=tuple(mutated_raw_values),
    )
    probes["slash_padding_projection_gate"] = _mutation_rejection(
        "three-row TSMIS slash-padding ledger",
        expected_padding,
        lambda: _padding_contract(_padding_artifacts(mutated_padding_rows)),
        mutated_input={
            "source_ref": padding_row.source_ref,
            "before": padding_row.raw_values[8],
            "after": mutated_raw_values[8],
        },
    )

    baseline_edition = _edition_row_contract(historical_excel, excel)
    _require(baseline_edition["source_updates"] == 6, "baseline edition gate drift")
    changed_identity = tuple(
        baseline_edition["paired_description_updates"][0]["identity"]
    )
    historical_index = next(
        index for index, row in enumerate(historical_excel)
        if (row.route, row.county, row.pm_base, 1) == changed_identity
    )
    current_match = next(
        row for row in excel
        if (row.route, row.county, row.pm_base, 1) == changed_identity
    )
    mutated_historical = list(historical_excel)
    old_row = mutated_historical[historical_index]
    mutated_historical[historical_index] = replace(
        old_row,
        values=(*old_row.values[:-1], current_match.values[-1]),
    )
    probes["historical_six_update_gate"] = _mutation_rejection(
        "six-update historical/current edition ledger",
        baseline_edition,
        lambda: _edition_row_contract(mutated_historical, excel),
        mutated_input={
            "source_ref": old_row.source_ref,
            "before": old_row.values[-1],
            "after": current_match.values[-1],
        },
    )

    mutated_documents = [list(records) for records in document_records]
    selected_document = next(
        index for index, records in enumerate(mutated_documents)
        if any(record["kind"] == "equate" for record in records)
    )
    equate_index = next(
        index for index, record in enumerate(mutated_documents[selected_document])
        if record["kind"] == "equate"
    )
    annotation = mutated_documents[selected_document][equate_index]
    following = mutated_documents[selected_document][equate_index + 1]
    mutated_documents[selected_document][equate_index], (
        mutated_documents[selected_document][equate_index + 1]
    ) = following, annotation
    expected_topology = {
        "equate_events": 998,
        "all_immediately_followed_by_owned_e_row": True,
    }
    probes["raw_equate_adjacency_and_order_gate"] = _mutation_rejection(
        "raw EQUATES immediate-following topology",
        expected_topology,
        lambda: {
            key: _raw_event_topology(mutated_documents)[key]
            for key in expected_topology
        },
        mutated_input={
            "annotation": {
                "member": annotation["member"], "route": annotation["route"],
                "pm": annotation["pm"], "line": annotation["line"],
            },
            "following": {
                "member": following["member"], "route": following["route"],
                "pm": following["pm"], "line": following["line"],
            },
            "mutation": "swap annotation and following row",
        },
    )

    escaped = [label for label, item in probes.items() if item["rejected"] is not True]
    return {
        "probe_count": len(probes),
        "rejected_count": len(probes) - len(escaped),
        "all_rejected": not escaped,
        "escaped_mutations": escaped,
        "probes": probes,
    }


def _write_new_atomic(path: Path, payload: bytes) -> None:
    _require(not path.exists(), f"refusing to overwrite checkpoint: {path}")
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent,
    )
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, path)
    except BaseException:
        try:
            os.unlink(temporary_name)
        except FileNotFoundError:
            pass
        raise


def run(output_root: Path, workers: int) -> dict[str, object]:
    _require(workers >= 1, "workers must be positive")
    _require(not output_root.exists(), f"output root must be clean: {output_root}")
    _require(output_root.parent.is_dir(), f"output parent missing: {output_root.parent}")
    output_root.mkdir()

    _require(INCOMPLETE_R1_ROOT.is_dir(), "preserved incomplete r1 root is missing")
    incomplete_r1_entries = sorted(
        str(path.relative_to(INCOMPLETE_R1_ROOT)).replace("\\", "/")
        for path in INCOMPLETE_R1_ROOT.rglob("*")
    )
    _require(
        not incomplete_r1_entries
        and not (INCOMPLETE_R1_ROOT / RESULT_NAME).exists(),
        "incomplete r1 unexpectedly contains a source-core publication",
    )

    bindings_before = _bind_static_inputs()
    accepted_stage6 = json.loads(STAGE6_RESULT.read_text(encoding="utf-8"))
    accepted_decision = json.loads(STAGE6_ACCEPTANCE.read_text(encoding="utf-8"))
    _validate_stage6_acceptance(accepted_stage6, accepted_decision)

    capture_binding = bindings_before["capture"]
    current_excel_capture, current_excel_capture_summary = _capture_tree_payloads(
        "current_tsmis_excel", ".xlsx", capture_binding,
    )
    historical_excel_capture, historical_excel_capture_summary = (
        _capture_tree_payloads(
            "historical_tsmis_excel_7_8", ".xlsx", capture_binding,
        )
    )
    current_pdf_capture, current_pdf_capture_summary = _capture_tree_payloads(
        "current_tsmis_pdf", ".pdf", capture_binding,
    )
    historical_pdf_capture, historical_pdf_capture_summary = (
        _capture_tree_payloads(
            "historical_tsmis_pdf_7_9", ".pdf", capture_binding,
        )
    )
    current_excel_source, current_excel_diagnostic = _parse_excel_captures(
        current_excel_capture, "current_tsmis_excel",
    )
    historical_excel_source, historical_excel_diagnostic = _parse_excel_captures(
        historical_excel_capture, "historical_tsmis_excel_7_8",
    )
    current_pdf_source, current_pdf_diagnostic = _parse_pdf_captures(
        current_pdf_capture, workers, "current_tsmis_pdf",
    )
    historical_pdf_source, historical_pdf_diagnostic = _parse_pdf_captures(
        historical_pdf_capture, workers, "historical_tsmis_pdf_7_9",
    )
    del (
        current_excel_capture, historical_excel_capture,
        current_pdf_capture, historical_pdf_capture,
    )

    _require(
        len(current_excel_source) == 60_494
        and len(historical_excel_source) == 60_493
        and len(current_pdf_source) == 60_493
        and len(historical_pdf_source) == 60_493,
        "TSMIS source census drift",
    )
    _require(
        {row.source for row in current_excel_source} == {"current_tsmis_excel"}
        and {row.source for row in historical_excel_source}
        == {"historical_tsmis_excel_7_8"}
        and {row.source for row in current_pdf_source} == {"current_tsmis_pdf"}
        and {row.source for row in historical_pdf_source}
        == {"historical_tsmis_pdf_7_9"},
        "typed current/historical TSMIS source provenance drift",
    )
    _require(
        current_excel_diagnostic["members"] == 252
        and historical_excel_diagnostic["members"] == 252
        and current_pdf_diagnostic["members"] == 252
        and historical_pdf_diagnostic["members"] == 252
        and current_pdf_diagnostic["pages"] == 3_177
        and current_pdf_diagnostic["data_pages"] == 2_673
        and current_pdf_diagnostic["description_fragments"] == 312
        and historical_pdf_diagnostic["pages"] == 3_177
        and historical_pdf_diagnostic["data_pages"] == 2_673
        and historical_pdf_diagnostic["description_fragments"] == 312,
        "TSMIS physical source accounting drift",
    )

    raw_records, raw_documents, document_records, raw_capture_summary = (
        _parse_raw_tsn()
    )
    normalized_sheet = _read_normalized_tsn()
    _require(
        normalized_sheet.pre_identity == normalized_sheet.post_identity,
        "accepted normalized workbook changed during streaming read",
    )
    fresh_stage6 = _fresh_stage6_proof(
        raw_records, raw_documents, normalized_sheet, accepted_stage6,
    )
    raw_event_topology = _raw_event_topology(document_records)
    _require(
        raw_event_topology["equate_events"] == 998
        and raw_event_topology["all_immediately_followed_by_owned_e_row"] is True,
        "raw TSN equate-event topology drift",
    )

    current_excel = _tsmis_rows(current_excel_source, "current_tsmis_excel")
    historical_excel = _tsmis_rows(
        historical_excel_source, "historical_tsmis_excel_7_8",
    )
    current_pdf = _tsmis_rows(current_pdf_source, "current_tsmis_pdf")
    historical_pdf = _tsmis_rows(
        historical_pdf_source, "historical_tsmis_pdf_7_9",
    )
    raw_tsn, raw_unknown = _raw_tsn_rows(raw_records)
    normalized_tsn = _normalized_tsn_rows(normalized_sheet.rows)
    _require(
        len(raw_tsn) == 69_758 and len(raw_unknown) == 46
        and len(normalized_tsn) == 69_758,
        "TSN comparison-row census drift",
    )

    edition = _edition_proof(
        historical_excel_source, current_excel_source,
        historical_pdf_source, current_pdf_source,
    )

    current_legs: dict[str, dict[str, object]] = {}
    current_pairs: dict[str, list[tuple[Row, Row]]] = {}
    keyable_raw_semantic_contracts: dict[str, object] = {}
    leg_specs = {
        "excel_vs_raw_tsn": (
            current_excel, raw_tsn,
            "current TSMIS Excel vs authoritative raw TSN",
            "full_pm_vs_tsn",
        ),
        "excel_vs_normalized_tsn": (
            current_excel, normalized_tsn,
            "current TSMIS Excel vs accepted normalized TSN",
            "full_pm_vs_tsn",
        ),
        "pdf_vs_raw_tsn": (
            current_pdf, raw_tsn,
            "current TSMIS PDF vs authoritative raw TSN",
            "full_pm_vs_tsn",
        ),
        "pdf_vs_normalized_tsn": (
            current_pdf, normalized_tsn,
            "current TSMIS PDF vs accepted normalized TSN",
            "full_pm_vs_tsn",
        ),
        "pdf_vs_excel": (
            current_pdf, current_excel,
            "current TSMIS PDF vs current TSMIS Excel",
            "same_source_base_pm",
        ),
    }
    for label, (left, right, description, policy) in leg_specs.items():
        keyable_result, pairs = _comparison(left, right, description, policy)
        result = keyable_result
        if label in ("excel_vs_raw_tsn", "pdf_vs_raw_tsn"):
            result = _complete_raw_publication_leg(keyable_result, raw_unknown)
            keyable_raw_semantic_contracts[label] = {
                "right_rows": keyable_result["right_rows"],
                "paired_rows": keyable_result["paired_rows"],
                "left_only_rows": keyable_result["left_only_rows"],
                "right_only_rows": keyable_result["right_only_rows"],
                "all_field_differing_rows": keyable_result[
                    "all_field_differing_rows"
                ],
                "all_field_difference_cells": keyable_result[
                    "all_field_difference_cells"
                ],
                "meaning": (
                    "semantic pairing over the 69,758 raw rows with printed County; "
                    "the complete publication contract is stored in current_source_legs"
                ),
            }
        _check_expected_leg(label, result, EXPECTED_CURRENT_LEGS[label])
        current_legs[label] = result
        current_pairs[label] = pairs

    same_source = current_legs["pdf_vs_excel"]
    _require(
        same_source["pairing"]["duplicate_groups"] == 4_030
        and tuple(same_source["pairing"]["assignment_cost"])
        == (3_721, 13_877, 0)
        and same_source["pairing"]["duplicate_trace_sha256"]
        == "919f7c60c40b0f4c94bb72f591e10af0ee3e0b663a7c54ccbde5b8bd98920d78",
        "same-source exact-DP assignment trace drift",
    )
    source_omissions = _same_source_omissions(same_source)
    source_events = _same_source_events(current_excel, current_pdf)
    _require(
        source_events["equate_events"] == 1_129
        and source_events["classes"] == {
            "pdf_only_suffix_on_following": 5,
            "same_convention": 852,
            "suffix_moved_between_annotation_and_following": 272,
        }
        and len(source_events["unlinked_pdf_e_rows"]) == 3
        and source_events["event_ledger_sha256"]
        == "7aea08e5b366dd4b163cadb5e713be832b885341a582f901fed3d6dca2e5d85a"
        and [
            item["excel_suffix"]
            for item in source_events["route_152_scr_t003_273_occurrences"]
        ] == ["", "E", ""]
        and [
            item["pdf_suffix"]
            for item in source_events["route_152_scr_t003_273_occurrences"]
        ] == ["", "", "E"],
        "current same-source EQUATES contract drift",
    )

    historical_legs: dict[str, dict[str, object]] = {}
    for label, right in (
        ("historical_excel_vs_raw_tsn", raw_tsn),
        ("historical_excel_vs_normalized_tsn", normalized_tsn),
    ):
        keyable_result, _pairs = _comparison(
            historical_excel, right,
            label.replace("_", " "), "full_pm_vs_tsn",
        )
        result = keyable_result
        if label == "historical_excel_vs_raw_tsn":
            result = _complete_raw_publication_leg(keyable_result, raw_unknown)
        expected = EXPECTED_HISTORICAL_LEGS[label]
        _check_expected_leg(label, result, expected)
        _require(
            tuple(result["pairing"]["assignment_cost"])
            == expected["assignment_cost"]
            and result["pairing"]["duplicate_trace_sha256"]
            == expected["duplicate_trace_sha256"],
            f"{label} typed historical cost/trace drift",
        )
        historical_legs[label] = result

    renamed_historical = [
        replace(row, dataset="renamed_without_semantic_role")
        for row in historical_excel
    ]
    rename_invariance: dict[str, object] = {}
    for label, right in (
        ("historical_excel_vs_raw_tsn", raw_tsn),
        ("historical_excel_vs_normalized_tsn", normalized_tsn),
    ):
        _pairs, _left_only, _right_only, pairing = _pair(
            renamed_historical,
            right,
            lambda row: (row.route, row.county, row.pm_full),
            _vs_tsn_values,
        )
        baseline = historical_legs[label]["pairing"]
        rename_invariance[label] = {
            "source_role": "tsmis",
            "source_format": "xlsx",
            "renamed_dataset": "renamed_without_semantic_role",
            "assignment_cost": pairing["assignment_cost"],
            "duplicate_trace_sha256": pairing["duplicate_trace_sha256"],
            "exact": (
                pairing["assignment_cost"] == baseline["assignment_cost"]
                and pairing["duplicate_trace_sha256"]
                == baseline["duplicate_trace_sha256"]
            ),
        }
    _require(
        all(item["exact"] for item in rename_invariance.values()),
        "historical dataset rename changed typed-role pairing",
    )

    raw_normalized = _raw_vs_normalized(raw_tsn, normalized_tsn)
    _require(
        raw_normalized["rows"] == 69_758
        and raw_normalized["differing_rows"] == 566
        and raw_normalized["difference_cells"] == 566
        and raw_normalized["field_difference_counts"] == {
            "Description": 1, "Distance To Next Point": 565,
        },
        "raw-vs-normalized semantic residue drift",
    )

    prefix_raw = _prefix_population(raw_tsn, "authoritative raw TSN")
    prefix_normalized = _prefix_population(
        normalized_tsn, "accepted normalized TSN",
    )
    for prefix_result in (prefix_raw, prefix_normalized):
        _require(
            prefix_result["rows"] == 154
            and prefix_result["owning_route_prefixes"] == 108
            and prefix_result["cross_route_prefixes"] == 46
            and prefix_result["changed_by_product"] == 154
            and prefix_result["numeric_prefixes_remaining_after_product"] == 2,
            "numeric Description prefix population drift",
        )
    _require(
        prefix_raw["context_multiset"] == prefix_normalized["context_multiset"]
        and prefix_raw["description_multiset"]
        == prefix_normalized["description_multiset"],
        "raw/normalized numeric-prefix population drift",
    )
    collapsed_raw = _collapsed_duplicate_distinctions(raw_tsn)
    collapsed_normalized = _collapsed_duplicate_distinctions(normalized_tsn)
    _require(
        collapsed_raw == collapsed_normalized
        and tuple(tuple(item["identity"]) for item in collapsed_raw)
        == EXPECTED_COLLAPSED_DUPLICATE_KEYS,
        "product-collapsed duplicate distinction drift",
    )

    padding = {
        "current_tsmis_excel": _padding_artifacts(current_excel),
        "current_tsmis_pdf": _padding_artifacts(current_pdf),
        "historical_tsmis_excel_7_8": _padding_artifacts(historical_excel),
        "historical_tsmis_pdf_7_9": _padding_artifacts(historical_pdf),
    }
    for label, artifacts in padding.items():
        _require(
            tuple(tuple(item["identity"]) for item in artifacts)
            == EXPECTED_PADDING_KEYS,
            f"{label} slash-padding census drift",
        )

    false_clean = {
        label: _false_clean_ledger(current_pairs[label])
        for label in (
            "excel_vs_raw_tsn", "excel_vs_normalized_tsn",
            "pdf_vs_raw_tsn", "pdf_vs_normalized_tsn",
        )
    }
    for label, ledger in false_clean.items():
        _require(
            ledger["rows"] == 81
            and ledger["wire_bytes"] == EXPECTED_PREFIX_LEDGER["wire_bytes"]
            and ledger["sha256"] == EXPECTED_PREFIX_LEDGER["sha256"]
            and ledger["content_sorted_sha256"]
            == EXPECTED_PREFIX_LEDGER["content_sorted_sha256"],
            f"{label} 81-row false-clean ledger drift",
        )

    mutations = _negative_mutations(
        current_excel, current_pdf, historical_pdf, historical_excel,
        raw_tsn, raw_unknown, document_records,
        source_events, prefix_raw, padding,
    )
    _require(mutations["all_rejected"], "a source-core negative mutation escaped")

    product_modules = [
        name for name in sys.modules
        if name == "compare_highway_sequence_tsn"
        or name.endswith(".compare_highway_sequence_tsn")
        or name == "compare_highway_sequence_pdf"
        or name.endswith(".compare_highway_sequence_pdf")
    ]
    _require(not product_modules, f"product comparison code was imported: {product_modules}")

    bindings_after = _bind_static_inputs()
    _require(
        bindings_after == bindings_before,
        "an exact source/capture/dependency/code identity changed during execution",
    )

    result: dict[str, object] = {
        "schema_version": 1,
        "audit": "Stage 8 Highway Sequence source-core comparison checkpoint",
        "artifact_status": "SOURCE_CORE_CHECKPOINT_NOT_FINAL_ACCEPTANCE",
        "acceptance_eligible": False,
        "stage8_family_accepted": False,
        "remaining_external_layers": [
            "product publication witness binding",
            "Comparison workbook and sidecar semantic inspection",
            "exhaustive evidence/source locator reconciliation",
            "permanent adversarial gate",
            "detached acceptance decision",
            "two byte-identical full replays",
        ],
        "prior_incomplete_attempts": [{
            "root": str(INCOMPLETE_R1_ROOT.resolve()),
            "entries": incomplete_r1_entries,
            "terminal_result_exists": False,
            "disposition": (
                "terminated before publication after adversarial audit-design review; "
                "not evidence and never eligible for acceptance"
            ),
        }],
        "independence": {
            "development_row_caches_read": False,
            "development_draft_results_read": False,
            "product_parsers_imported": False,
            "product_comparators_imported": False,
            "product_evidence_imported": False,
            "tsmis_parser": (
                "audit-owned positional Excel plus fixed-grid pdfplumber source oracle"
            ),
            "tsn_parser": "accepted audit-owned Stage-6 raw-PDF parser",
            "normalized_reader": "generic audit-owned stdlib OOXML stream reader",
        },
        "bindings_before": bindings_before,
        "bindings_after": bindings_after,
        "bindings_stable": True,
        "accepted_stage6_chain": {
            "decision": accepted_decision["decision"],
            "required_result_flags": accepted_decision["required_result_flags"],
            "tracked_identities": accepted_decision["tracked_identities"],
            "fresh_reparse": fresh_stage6,
        },
        "source_datasets": {
            "current_tsmis_excel": _diagnostic_summary(
                current_excel_source, current_excel_diagnostic,
            ),
            "historical_tsmis_excel_7_8": _diagnostic_summary(
                historical_excel_source, historical_excel_diagnostic,
            ),
            "current_tsmis_pdf": _diagnostic_summary(
                current_pdf_source, current_pdf_diagnostic,
            ),
            "historical_tsmis_pdf_7_9": _diagnostic_summary(
                historical_pdf_source, historical_pdf_diagnostic,
            ),
            "authoritative_raw_tsn": {
                "records": len(raw_records),
                "data_rows": sum(
                    record["kind"] == "data" for record in raw_records
                ),
                "equates": sum(
                    record["kind"] == "equate" for record in raw_records
                ),
                "known_county_rows": len(raw_tsn),
                "unknown_county_equates": len(raw_unknown),
                "members": len(raw_documents),
                "pages": sum(document["page_count"] for document in raw_documents),
                "single_byte_capture": raw_capture_summary,
            },
            "accepted_normalized_tsn": {
                "rows": len(normalized_tsn),
                "columns": 8,
                "pre_identity": asdict(normalized_sheet.pre_identity),
                "post_identity": asdict(normalized_sheet.post_identity),
            },
        },
        "single_byte_tsmis_captures": {
            "current_tsmis_excel": current_excel_capture_summary,
            "historical_tsmis_excel_7_8": historical_excel_capture_summary,
            "current_tsmis_pdf": current_pdf_capture_summary,
            "historical_tsmis_pdf_7_9": historical_pdf_capture_summary,
            "all_parsers_consumed_the_exact_bound_payload_bytes": True,
            "filesystem_pre_post_manifests_are_mutation_guards_only": True,
        },
        "edition_proof": edition,
        "same_source_equate_events": source_events,
        "raw_tsn_equate_events": raw_event_topology,
        "same_source_unrepresented_claims": source_omissions,
        "raw_vs_normalized_tsn": raw_normalized,
        "current_source_legs": current_legs,
        "keyable_raw_semantic_contracts": keyable_raw_semantic_contracts,
        "historical_regression_legs": historical_legs,
        "typed_role_rename_invariance": rename_invariance,
        "description_prefix_proof": {
            "modeled_product_rule": {
                "module_identity": bindings_before["files"][
                    "modeled_product_comparator"
                ],
                "regex": ROUTE_PREFIX_RE.pattern,
                "product_code_was_not_imported": True,
            },
            "raw_population": prefix_raw,
            "normalized_population": prefix_normalized,
            "collapsed_duplicate_distinctions": collapsed_raw,
            "false_clean_ledgers": false_clean,
        },
        "tsmis_slash_padding_artifacts": padding,
        "negative_mutations": mutations,
        "source_core_invariants": {
            "current_rows_excel_pdf": [60_494, 60_493],
            "same_source_shape_pdf_excel": [60_493, 0, 1],
            "same_source_differences": {
                "rows": 1_410,
                "cells": 3_721,
                "fields": EXPECTED_CURRENT_LEGS["pdf_vs_excel"]["all"][2],
            },
            "same_source_equate_events": 1_129,
            "same_source_equate_classes": source_events["classes"],
            "current_vs_historical_updates": 6,
            "raw_tsn_records": 69_804,
            "raw_tsn_data_plus_equates": [68_806, 998],
            "raw_tsn_unknown_county_equates": 46,
            "complete_raw_publication_shapes": {
                "excel_vs_raw_tsn": [57_072, 3_422, 12_732],
                "pdf_vs_raw_tsn": [57_505, 2_988, 12_299],
            },
            "keyable_raw_semantic_shapes": {
                "excel_vs_raw_tsn": [57_072, 3_422, 12_686],
                "pdf_vs_raw_tsn": [57_505, 2_988, 12_253],
            },
            "normalized_tsn_rows": 69_758,
            "pointer_tokens_blanked": 565,
            "wrapped_description_punctuation_delta": 1,
            "numeric_prefix_descriptions_each_tsn_form": 154,
            "product_false_clean_rows_each_current_tsn_leg": 81,
            "slash_padding_rows_each_tsmis_form": 3,
            "corrected_current_leg_contracts": 5,
            "historical_typed_role_contracts": 2,
            "all_checks_passed": True,
        },
    }
    payload = _json_bytes(result)
    output_path = output_root / RESULT_NAME
    _write_new_atomic(output_path, payload)
    reopened = output_path.read_bytes()
    _require(reopened == payload, "published source-core bytes failed exact reopen")
    print(json.dumps({
        "status": "PASS_SOURCE_CORE_CHECKPOINT_NOT_FINAL_ACCEPTANCE",
        "output": str(output_path.resolve()),
        "bytes": len(payload),
        "sha256": _sha_bytes(payload),
        "current_source_legs": 5,
        "source_core_checks_passed": True,
    }, sort_keys=True))
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-root", type=Path, required=True,
        help="Required nonexistent directory for this source-core checkpoint.",
    )
    parser.add_argument(
        "--workers", type=int, default=min(8, os.cpu_count() or 1),
    )
    arguments = parser.parse_args()
    run(arguments.output_root.resolve(), arguments.workers)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (SourceCoreError, tsmis_source.OracleError, stage6.ConservationError) as error:
        print(f"FAIL Highway Sequence source-core checkpoint: {error}")
        raise SystemExit(1)
