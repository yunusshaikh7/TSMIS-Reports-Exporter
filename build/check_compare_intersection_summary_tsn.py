"""Golden check for the TSMIS-vs-TSN Intersection Summary comparator
(scripts/compare_intersection_summary_tsn.py) — the AGGREGATE recipe applied to
the intersection taxonomy.

Locks: the canonical spec (66 categories after the signal fold; unique keys/slugs),
the CONTROL-TYPE signal crosswalk (TSN sub-types J–P fold into the shared 'S -
SIGNALIZED' category, matching the Detail — so Signalized compares on BOTH sides and
there are NO TSN-only codes left), the remaining genuinely one-sided TSMIS-only codes
(Roundabout R / PHB O / Flash Q / intersection R/C/P / left-chan Y — absent from the
TSN summary PDF), the '+ no data' buckets the TSN PDF reports as 0 being compared
(num-of-lanes '+' is now BOTH), the spec-driven block-walk mapper (incl. the J–P→S
fold and the Rural/Urban '-O OUTSIDE CITY' parent disambiguation — a count-less
parent still binds its child and a counted orphan '-O' refuses, CMP-AUD-023), the
strict consolidated-workbook summing loader (CMP-AUD-021/022: numeric text parses,
fractions/booleans/duplicates refuse, absent columns stay absent), the independent
per-side partition validation with exposed bounded residuals (CMP-AUD-020), and
end-to-end structure. No Excel; CI-safe.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_intersection_summary_tsn.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_intersection_summary_tsn as cmp
import summary_layout as sl
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []
SPEC = sl.INTERSECTION_SUMMARY_SPEC
_KEY = {slug: key for key, slug in cmp._CATEGORIES}


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def test_spec():
    print("spec + signal fold + one-sided marking:")
    sc = cmp._SCHEMA
    check("header Category/Count, key_field 0", sc.header == ["Category", "Count"] and sc.key_field == 0)
    check("extra_sheet_writer set", sc.extra_sheet_writer is not None)
    keys = [k for k, _ in cmp._CATEGORIES]
    slugs = [s for _, s in cmp._CATEGORIES]
    check("66 categories (J-P folded into Signalized), unique keys + slugs",
          len(cmp._CATEGORIES) == 66 and len(set(keys)) == 66 and len(set(slugs)) == 66)
    by_slug = {c.slug: c for sec in SPEC.sections for c in sec.cats}
    check("CONTROL signal sub-types J-P are NOT separate categories (folded -> S)",
          not any(f"is_control_types_{c.lower()}" in by_slug for c in "JKLMNP"))
    check("CONTROL S - Signalized is now BOTH (compared, not one-sided)",
          by_slug["is_control_types_s"].sides == "both")
    check("CONTROL O/Q/R stay TSMIS-only (absent from the TSN summary)",
          all(by_slug[f"is_control_types_{c.lower()}"].sides == "tsmis" for c in "OQR"))
    check("INTERSECTION TYPE Roundabout (R) is TSMIS-only",
          by_slug["is_intersection_type_r"].sides == "tsmis")
    check("MAINLINE NUM OF LANES '+' is now BOTH (TSN reports it as 0)",
          by_slug["is_mainline_num_of_lanes_plus"].sides == "both")
    check("no TSN-only categories remain (signals folded)", len(sl._IS_TSN_ONLY) == 0)
    tsn_cats = dict(SPEC.categories_for("tsn"))
    check("categories_for(tsn) now INCLUDES Signalized (S)", _KEY["is_control_types_s"] in tsn_cats)


def test_block_walk():
    print("spec-driven block-walk mapper (counts_from_rows):")
    rows = [
        (None, "HIGHWAY GROUP"), (None, "NUMBER CODE"),
        (5, "U-UNDIVIDED"), (3, "D-DIVIDED"),
        (None, "<---RURAL/URBAN/SUBURBAN--->"), (None, "NUMBER CODE"),
        (1, "R-RURAL -I INSIDE CITY"), (9, "-O OUTSIDE CITY"),
        (2, "U-URBAN -I INSIDE CITY"), (4, "-O OUTSIDE CITY"),
        (None, "CONTROL TYPES"), (7, "A-NO CONTROL"), (6, "S-SIGNALIZED"),
        (2, "J-SIGNAL PRETIMED (2-PHASE)"), (3, "P-SIGNALS FULL-ACTUATED (MULTI-PHASE)"),
        (None, "MAINLINE NUM OF LANES"), (8, "2"), (1, "+-NO DATA GIVEN"),
    ]
    c = sl.counts_from_rows(SPEC, rows)
    check("HG undivided=5, divided=3", c["is_highway_group_u"] == 5 and c["is_highway_group_d"] == 3)
    check("rural '-O' bound to R-RURAL parent (R-O=9)", c["is_rural_urban_suburban_r_o"] == 9)
    check("urban '-O' bound to U-URBAN parent (U-O=4)", c["is_rural_urban_suburban_u_o"] == 4)
    # CROSSWALK: J–P fold into S — Signalized accumulates 6 (S) + 2 (J) + 3 (P) = 11.
    check("control A=7; signals fold into S (S 6 + J 2 + P 3 = 11)",
          c["is_control_types_a"] == 7 and c["is_control_types_s"] == 11)
    check("folded sub-types leave no separate J/P slug",
          "is_control_types_j" not in c and "is_control_types_p" not in c)
    check("lanes '2'=8 and '+'=1", c["is_mainline_num_of_lanes_2"] == 8 and c["is_mainline_num_of_lanes_plus"] == 1)

    # CMP-AUD-023: the Rural/Urban parent updates from the LABEL even on a
    # count-less row, so the following '-O' binds to the RIGHT parent…
    ru = sl._IS_RURAL_URBAN
    c = sl.counts_from_rows(SPEC, [(None, ru), (None, "U-URBAN -I INSIDE CITY"),
                                   (4, "-O OUTSIDE CITY")])
    check("count-less U-URBAN parent still binds '-O' to U-O (never R-O)",
          c.get("is_rural_urban_suburban_u_o") == 4
          and "is_rural_urban_suburban_r_o" not in c)
    # …and a COUNTED orphan '-O' (no parent at all) is an error, never Rural.
    try:
        sl.counts_from_rows(SPEC, [(None, ru), (5, "-O OUTSIDE CITY")])
        check("counted orphan '-O' refuses (never defaults to Rural)", False)
    except ValueError as e:
        check("counted orphan '-O' refuses (never defaults to Rural)",
              "no preceding" in str(e))
    check("count-less orphan '-O' text is ignored (no count to misfile)",
          sl.counts_from_rows(SPEC, [(None, ru), (None, "-O OUTSIDE CITY")]) == {})
    # CMP-AUD-021: strict counts in the shared mapper.
    try:
        sl.counts_from_rows(SPEC, [(None, "HIGHWAY GROUP"), (1.5, "U-UNDIVIDED")])
        check("fractional count in the block-walk refuses", False)
    except ValueError as e:
        check("fractional count in the block-walk refuses", "fractional" in str(e))


def _write_tsmis(path, cols, route_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = cmp.TSMIS_SHEET
    ws.append(["Route", "Total Intersections"] + cols)
    for r in route_rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _write_tsn(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = cmp.NORMALIZED_SHEET
    ws.append(["Category", "Count"])
    for k, v in rows:
        ws.append([k, v])
    wb.save(path)
    wb.close()


def _sheet(path, name):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
        return header, rows
    finally:
        wb.close()


# Fixture builders: arithmetically CONSISTENT tables per the censused partition
# contract (CMP-AUD-020) — the TSMIS side partitions every block to the total
# except the bounded Highway Group; the TSN side partitions its exact blocks and
# may run short only in the bounded ones.
_ALL_KEYS = [c.key for sec in SPEC.sections for c in sec.cats]


def _tsmis_table(total, values):
    """One consolidated route row carrying {slug: count} (unlisted slugs 0)."""
    return ["001", total] + [values.get(c.slug, 0)
                             for sec in SPEC.sections for c in sec.cats]


def _tsn_table(values):
    """The normalized TSN rows for {slug: count} (unlisted TSN-applicable slugs 0)."""
    return [(key, values.get(slug, 0)) for key, slug in SPEC.categories_for("tsn")]


def _consistent_tsmis(total=10):
    """HG U=8 (bounded: site under-count, residual 2); every other block == total,
    with the original flavors kept: control A=3 / S=5 / R=2 (R is TSMIS-only)."""
    return {
        "is_highway_group_u": 8,
        "is_rural_urban_suburban_r": total,
        "is_intersection_type_f": total,
        "is_lighting_type_n": total,
        "is_control_types_a": 3, "is_control_types_s": 5, "is_control_types_r": 2,
        "is_mainline_num_of_lanes_2": total,
        "is_mainline_mastarm_y": total,
        "is_mainline_left_channelization_n": total,
        "is_mainline_right_channelization_y": total,
        "is_mainline_traffic_flow_p": total,
    }


def _consistent_tsn(total=12):
    """HG exact (U=8 matching + D=4); intersection-type bounded SHORT by 1 (the
    censused-untabulated residual, exposed as a note); every exact block == total."""
    return {
        "is_highway_group_u": 8, "is_highway_group_d": 4,
        "is_rural_urban_suburban_r": total,
        "is_intersection_type_f": total - 1,
        "is_lighting_type_n": total,
        "is_control_types_a": 4, "is_control_types_s": 7, "is_control_types_z": 1,
        "is_mainline_num_of_lanes_2": total,
        "is_mainline_mastarm_y": total,
        "is_mainline_left_channelization_n": total,
        "is_mainline_right_channelization_y": total,
        "is_mainline_traffic_flow_p": total,
        "total_intersections": total,
    }


def test_end_to_end():
    print("end-to-end (signal fold; Signalized compared; only TSMIS-only codes one-sided):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_is_tsn_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # TSMIS total 10: U=8 (HG bounded), A=3, S=5, R=2 (TSMIS-only).
    _write_tsmis(tsmis, _ALL_KEYS, [_tsmis_table(10, _consistent_tsmis())])
    # TSN total 12: U matches (8), A differs (3 vs 4), S differs (5 vs 7, compared
    # both sides), total differs (10 vs 12); R has no TSN row -> Only in TSMIS.
    _write_tsn(tsn, _tsn_table(_consistent_tsn()))
    res = cmp.compare(tsmis, tsn, out, events=Events(), confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")

    header, rows = _sheet(out, "Comparison")
    si = header.index("Status")
    both = sum(1 for r in rows if r[si] == "Both")
    tonly = sum(1 for r in rows if r[si] == "TSMIS only")
    nonly = sum(1 for r in rows if r[si] == "TSN only")
    check("structural split: 58 both / 8 only-TSMIS / 0 only-TSN (signals folded)",
          both == 58 and tonly == 8 and nonly == 0)

    oa = {r[0] for r in _sheet(out, "Only in TSMIS")[1]}
    check("no TSN-only categories (signals fold into shared Signalized)", nonly == 0)
    check("TSMIS-only Roundabout (R) is in 'Only in TSMIS'", _KEY["is_control_types_r"] in oa)

    cat_col, cnt_col = header.index("Category"), header.index("Count")
    by = {r[cat_col]: r for r in rows}
    DIFF = " ≠ "
    check("matched HG Undivided (8) has no diff", DIFF not in by[_KEY["is_highway_group_u"]][cnt_col])
    check("Signalized (S) is COMPARED on both and differs (5 vs 7)",
          DIFF in by[_KEY["is_control_types_s"]][cnt_col])
    check("control A differs (3 vs 4)", DIFF in by[_KEY["is_control_types_a"]][cnt_col])
    check("Total Intersections differs (10 vs 12)", DIFF in by[SPEC.total.key][cnt_col])

    fh, fr = _sheet(out, SPEC.sheet_name)
    flat = [c for row in [fh] + fr for c in row]
    check("familiar sheet labels TSMIS/TSN", "TSMIS" in flat and "TSN" in flat)
    check("familiar sheet shows the folded Signalized category (incl. TSN J-P)",
          any("SIGNALIZED" in c for c in flat) and any("J-P" in c or "J–P" in c for c in flat))
    # CMP-AUD-020: the bounded residuals are EXPOSED as familiar-sheet notes —
    # the TSMIS Highway Group under-count (8 of 10) and the TSN-untabulated
    # intersection-type remainder (11 of 12) — never fabricated into a category,
    # never a warning.
    check("familiar sheet exposes the TSMIS Highway Group residual note (2 not)",
          any("HIGHWAY GROUP" in c and "2 not" in c and "TSMIS" in c for c in flat))
    check("familiar sheet exposes the TSN intersection-type residual note (1 not)",
          any("INTERSECTION TYPE" in c and "1 not" in c and "TSN" in c for c in flat))
    # CMP-AUD-184: the note must describe the cells truthfully (structural
    # absence stays BLANK, distinguished from an explicit source 0) and must not
    # cite another family's categories on this sheet.
    check("familiar note says one-sided categories stay BLANK (no zero-fill claim)",
          any("stays BLANK" in c and "real source zero" in c for c in flat)
          and not any("show 0 on that side" in c for c in flat))
    check("familiar note no longer cites Ramp P/V on the Intersection sheet",
          not any("ramp types P / V" in c for c in flat))
    print(f"      (both={both}, only-TSMIS={tonly}, only-TSN={nonly})")


def test_one_sided_familiar_agreement():
    """CMP-AUD-184 mutation sweep: EVERY structurally one-sided Intersection
    category must agree across the familiar sheet (value on the classifying
    side, BLANK absent side, BLANK Δ) and the generic Comparison status
    ('Only in TSMIS'), in BOTH the formulas and values workbooks."""
    print("one-sided familiar/Comparison agreement, both modes (CMP-AUD-184):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_is_184_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # Give every TSMIS-only category a non-zero count, inside blocks that still
    # reconcile: control 1+2+3+2+2=10; int-type 1+2+3+1+3=10; left-chan 2+8=10.
    vals = _consistent_tsmis()
    vals.update({
        "is_control_types_a": 1, "is_control_types_s": 2,
        "is_control_types_r": 3, "is_control_types_o": 2, "is_control_types_q": 2,
        "is_intersection_type_f": 1, "is_intersection_type_r": 2,
        "is_intersection_type_c": 3, "is_intersection_type_p": 1,
        "is_intersection_type_plus": 3,
        "is_mainline_left_channelization_y": 2,
        "is_mainline_left_channelization_n": 8,
    })
    _write_tsmis(tsmis, _ALL_KEYS, [_tsmis_table(10, vals)])
    _write_tsn(tsn, _tsn_table(_consistent_tsn()))
    res = cmp.compare(tsmis, tsn, out, events=Events(),
                      confirm_overwrite=lambda _p: True, mode="both")
    check("compare ok (both modes)", res.status == "ok")
    one_sided = [(sec.name, c) for sec in SPEC.sections for c in sec.cats
                 if c.sides == "tsmis"]
    check("sweep covers all 8 TSMIS-only categories", len(one_sided) == 8)

    values_twin = out.with_name(f"{out.stem} (values){out.suffix}")
    for path, label in ((out, "formulas"), (values_twin, "values")):
        fh, fr = _sheet(path, SPEC.sheet_name)
        # Familiar rows keyed by (section, label): several sections share the
        # '+ - NO DATA GIVEN' label, so a flat label lookup is ambiguous.
        section_names = {s.name for s in SPEC.sections}
        by_sec_label, cur = {}, None
        for r in fr:
            if not r or not r[0]:
                continue
            if r[0] in section_names:
                cur = r[0]
            elif cur is not None:
                by_sec_label[(cur, r[0])] = r
        agree = []
        for sec_name, c in one_sided:
            row = by_sec_label.get((sec_name, c.label))
            v = vals.get(c.slug, 0)
            agree.append(row is not None and row[1] == str(v)
                         and row[2] == "" and row[3] == "")
        check(f"{label}: every one-sided row shows value/BLANK/BLANK", all(agree))
    # Generic statuses read from the VALUES twin (the formulas twin's Status
    # cells are formulas, unreadable under data_only).
    header, rows = _sheet(values_twin, "Comparison")
    si, ci = header.index("Status"), header.index("Category")
    status_by_key = {r[ci]: r[si] for r in rows}
    check("generic Comparison marks all 8 'TSMIS only'",
          all(status_by_key.get(c.key) == "TSMIS only" for _s, c in one_sided))


def test_validation_refusals():
    """CMP-AUD-020/021/022: a table that does not reconcile refuses with a named
    block; missing categories/totals are hard stops; duplicate exact keys refuse
    (while DISTINCT stale J–P/S keys still fold)."""
    print("independent per-side validation (CMP-AUD-020/021/022):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_is_val_"))
    good_tsmis, good_tsn = root / "t.xlsx", root / "n.xlsx"
    _write_tsmis(good_tsmis, _ALL_KEYS, [_tsmis_table(10, _consistent_tsmis())])
    _write_tsn(good_tsn, _tsn_table(_consistent_tsn()))

    def refuses(label, tsmis, tsn, *needles):
        try:
            cmp._load_pair(tsmis, tsn)
            check(label, False)
        except ValueError as e:
            check(label, all(n in str(e) for n in needles))

    # All-zero categories under a non-zero total refuse, even both sides agreeing.
    zt = root / "t_zero.xlsx"
    _write_tsmis(zt, _ALL_KEYS, [_tsmis_table(10, {})])
    refuses("all-zero TSMIS categories + total=10 refuse naming a block",
            zt, good_tsn, "does not reconcile")
    zn = root / "n_zero.xlsx"
    _write_tsn(zn, _tsn_table({"total_intersections": 10}))
    refuses("all-zero TSN categories + total=10 refuse too",
            good_tsmis, zn, "does not reconcile")

    # A missing TSN category / total is a hard stop, never a fabricated 0.
    mn = root / "n_missing.xlsx"
    _write_tsn(mn, [(k, v) for k, v in _tsn_table(_consistent_tsn())
                    if k != _KEY["is_lighting_type_n"]])
    refuses("missing TSN category is a hard stop naming it",
            good_tsmis, mn, "N - NO LIGHTING", "missing")
    tn = root / "n_no_total.xlsx"
    _write_tsn(tn, [(k, v) for k, v in _tsn_table(_consistent_tsn())
                    if k != SPEC.total.key])
    refuses("missing TSN grand total is a hard stop",
            good_tsmis, tn, "grand total")

    # An over-total bounded block refuses (sums can only run SHORT).
    ot = root / "n_over.xlsx"
    over = _consistent_tsn()
    over["is_intersection_type_f"] = 99
    _write_tsn(ot, _tsn_table(over))
    refuses("a bounded block summing ABOVE the total refuses",
            good_tsmis, ot, "MORE than the total")

    # Duplicate exact keys refuse (was: silently summed, CMP-AUD-022)…
    dn = root / "n_dup.xlsx"
    _write_tsn(dn, [(_KEY["is_highway_group_u"], 4), (_KEY["is_highway_group_u"], 7)])
    try:
        cmp._load_tsn(dn)
        check("duplicate exact normalized key refuses (no summing)", False)
    except ValueError as e:
        check("duplicate exact normalized key refuses (no summing)", "twice" in str(e))
    # …including a REPEATED stale J key (only DISTINCT legacy keys may fold).
    dj = root / "n_dup_j.xlsx"
    _write_tsn(dj, [("CONTROL TYPES: J - SIGNAL PRETIMED (2-PHASE)", 1),
                    ("CONTROL TYPES: J - SIGNAL PRETIMED (2-PHASE)", 2)])
    try:
        cmp._load_tsn(dj)
        check("repeated stale J key refuses", False)
    except ValueError as e:
        check("repeated stale J key refuses", "twice" in str(e))

    # Strict counts: numeric text parses; fractions refuse (CMP-AUD-021).
    st = root / "n_text.xlsx"
    _write_tsn(st, [(_KEY["is_highway_group_u"], "1,234")])
    check("comma text '1,234' parses as 1234 (not dropped)",
          cmp._load_tsn(st)["is_highway_group_u"] == 1234)
    sf = root / "n_frac.xlsx"
    _write_tsn(sf, [(_KEY["is_highway_group_u"], 1.9)])
    try:
        cmp._load_tsn(sf)
        check("fractional normalized count refuses", False)
    except ValueError as e:
        check("fractional normalized count refuses", "fractional" in str(e))

    # A duplicated TSMIS category column refuses (CMP-AUD-022).
    dt = root / "t_dupcol.xlsx"
    _write_tsmis(dt, _ALL_KEYS + [_KEY["is_highway_group_u"]],
                 [_tsmis_table(10, _consistent_tsmis()) + [1]])
    try:
        cmp._load_tsmis(dt)
        check("duplicated TSMIS category column refuses", False)
    except ValueError as e:
        check("duplicated TSMIS category column refuses", "duplicated" in str(e))


def test_stale_library_fold():
    """A normalized library built BEFORE the signal fold (separate J–P category rows
    + the old 'S - SIGNALIZED' key) is REUSED, not rebuilt, after the code change
    (tsn_library.build_consolidated reuses when current). _load_tsn must fold those
    stale keys into Signalized on read, so a reused pre-fold library still compares
    correctly. Regression lock paired with the Detail's normalized-path repair."""
    print("stale-library signal fold (reused pre-fold normalized workbook):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_is_stale_"))
    stale = root / "stale.xlsx"
    _write_tsn(stale, [
        ("CONTROL TYPES: A - NO CONTROL", 1760),
        ("CONTROL TYPES: J - SIGNAL PRETIMED (2-PHASE)", 207),
        ("CONTROL TYPES: P - SIGNALS FULL-ACTUATED (MULTI-PHASE)", 2023),
        ("CONTROL TYPES: S - SIGNALIZED", 5),        # the OLD label/key
    ])
    rec = cmp._load_tsn(stale)
    check("stale J(207)+P(2023)+old-S(5) fold into Signalized (2235)",
          rec.get(cmp._SIGNALIZED_SLUG) == 207 + 2023 + 5)
    check("non-signal A unchanged (1760)", rec.get("is_control_types_a") == 1760)


def test_corrupt_pdf_is_valueerror():
    """A corrupt/truncated statewide PDF must honor the loader contract:
    ValueError, never a raw pdfplumber exception escaping into the matrix path."""
    import tempfile
    bad = Path(tempfile.mkdtemp()) / "TSN statewide.pdf"
    bad.write_bytes(b"%PDF-1.4 not really a pdf, just junk bytes with no xref")
    try:
        cmp._load_tsn(bad)
        check("corrupt PDF raises", False)
    except ValueError as e:
        check("corrupt PDF -> ValueError (loader contract)", True)
        check("...message names the file", "TSN statewide.pdf" in str(e))
    except Exception as e:  # noqa: BLE001 — the point of the test
        check(f"corrupt PDF -> ValueError, not {type(e).__name__}", False)


def main():
    test_spec()
    test_block_walk()
    test_end_to_end()
    test_validation_refusals()
    test_one_sided_familiar_agreement()
    test_stale_library_fold()
    test_corrupt_pdf_is_valueerror()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-INTERSECTION-SUMMARY-TSN CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
