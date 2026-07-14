#!/usr/bin/env python3
"""Audit both Highway Sequence product legs against the raw-TSN dev twin.

The product witnesses deliberately use a cache-derived 69,804-row workbook so
that current comparison behavior can be observed without silently dropping the
46 pre-county equates, blanking 565 pointer tokens, or accepting the one
description punctuation mutation in the normalized TSN workbook.  This checker
imports no product comparison, product sidecar reader, or product publisher.

It reuses only the already-proven, product-independent workbook/publication
primitives from ``check_phase8_highway_sequence_product_comparisons`` and adds
raw-twin-specific source, provenance, and claim-conservation checks.  A PASS is
therefore a faithful product witness, not source acceptance: the raw twin was
built from a development cache and final acceptance must reparse the immutable
TSN PDFs.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime, time
import hashlib
import json
from pathlib import Path
import re
from typing import Iterable, Sequence
import zipfile

from openpyxl import load_workbook

import check_phase8_highway_sequence_product_comparisons as base


VISUAL_ROOT = base.VISUAL_ROOT
DEFAULT_OUTPUT = (
    VISUAL_ROOT
    / "phase8_highway_sequence_product_raw_comparison_parity_dev_r1.json"
)

HEADERS = (
    "Route", "County", "PM", "City", "HG", "FT",
    "Distance To Next Point", "Description",
)
CURRENT_HEADERS = (
    "Route", "County", "City", None, "PM", None, "HG", "FT",
    "Distance To Next Point", "Description",
)

INPUT_BINDINGS = {
    "current_excel": {
        "bytes": 2_424_212,
        "sha256": "cf5905332db3d3eb5a49a87d603f6e36f209cad9a84173b381dace6600168b20",
    },
    "current_pdf": {
        "bytes": 2_371_547,
        "sha256": "070afe51ea3bf84c9704d0a36a02702b65189941badab6374b03461db8ef6ccc",
    },
    "raw_tsn_workbook": {
        "bytes": 2_541_734,
        "sha256": "d594e2441b81c4d4d81c11aa5bbf01418bcd2dcc0bedf3ee9a6221a66cb03fa1",
    },
    "raw_tsn_provenance": {
        "bytes": 23_610_997,
        "sha256": "f27c7724f9acc8988bfd65c896e8278853b70690ed36d0317fabf6c5af8920f2",
    },
    "raw_tsn_manifest": {
        "bytes": 6_464,
        "sha256": "c534e818d4c1aacf7a72ff4a623a4c9c53d8b5d32cb4d3d8b0fc19958e567533",
    },
    "raw_tsn_result": {
        "bytes": 2_428,
        "sha256": "51a0cfb70611442fc5b7ca4bb1acbb2779446b7d5400d10590d31c798629d1bc",
    },
}

RAW_COUNTS = {
    "raw_records": 69_804,
    "data_records": 68_806,
    "equate_records": 998,
    "pre_county_equates": 46,
    "projectable_records": 69_758,
    "pointer_P": 283,
    "pointer_arrow": 282,
    "pointer_total": 565,
}

LEG_SPECS = {
    "excel_vs_raw_tsn": {
        "side_a": "TSMIS", "side_b": "TSN",
        "side_a_input": "current_excel",
        "side_a_rows": 60_494, "side_b_rows": 69_804,
        "paired_rows": 57_072, "side_a_only_rows": 3_422,
        "side_b_only_rows": 12_732,
    },
    "pdf_vs_raw_tsn": {
        "side_a": "TSMIS (PDF)", "side_b": "TSN",
        "side_a_input": "current_pdf",
        "side_a_rows": 60_493, "side_b_rows": 69_804,
        "paired_rows": 57_505, "side_a_only_rows": 2_988,
        "side_b_only_rows": 12_299,
    },
}

_CONTROL_WS = re.compile(r"[\t\n\r\f\v]")
_DESC_PREFIX = re.compile(r"^\d{1,3}[A-Z]?/")


class AuditError(base.AuditError):
    """The raw-twin witness or one of its immutable bindings drifted."""


def _reject_constant(value: str) -> None:
    raise AuditError(f"non-finite JSON constant is forbidden: {value}")


def _unique_object(pairs: list[tuple[str, object]]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in pairs:
        if key in result:
            raise AuditError(f"duplicate JSON key is forbidden: {key!r}")
        result[key] = value
    return result


def _strict_canonical_json(
    path: Path, *, maximum: int | None = None,
) -> tuple[dict[str, object], dict[str, object]]:
    identity = base._identity(path)
    raw = path.read_bytes()
    if maximum is not None and len(raw) > maximum:
        raise AuditError(f"{path.name}: JSON exceeds {maximum:,} bytes")
    try:
        value = json.loads(
            raw.decode("utf-8"), object_pairs_hook=_unique_object,
            parse_constant=_reject_constant,
        )
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise AuditError(f"{path.name}: invalid strict UTF-8 JSON: {exc}") from exc
    if not isinstance(value, dict):
        raise AuditError(f"{path.name}: JSON root is not an object")
    if base._json_line(value) != raw:
        raise AuditError(f"{path.name}: JSON is not canonical/LF-terminated")
    return value, identity


def _embedded_identity_matches(
    value: object, identity: dict[str, object],
) -> bool:
    if not isinstance(value, dict):
        return False
    raw_path = value.get("canonical_path", value.get("path"))
    raw_size = value.get("bytes", value.get("size"))
    try:
        path = str(Path(str(raw_path)).resolve())
    except (OSError, RuntimeError, ValueError):
        return False
    return (
        path == identity["path"]
        and raw_size == identity["bytes"]
        and value.get("sha256") == identity["sha256"]
    )


def _authenticate_embedded_file(value: object, label: str) -> dict[str, object]:
    if not isinstance(value, dict):
        raise AuditError(f"{label}: embedded identity is not an object")
    raw_path = value.get("canonical_path", value.get("path"))
    if not isinstance(raw_path, str):
        raise AuditError(f"{label}: embedded path is absent")
    identity = base._identity(Path(raw_path).resolve())
    if not _embedded_identity_matches(value, identity):
        raise AuditError(f"{label}: embedded identity differs from current bytes")
    return identity


def _bind_inputs(record: dict[str, object], label: str) -> dict[str, dict[str, object]]:
    inputs = record.get("inputs")
    if not isinstance(inputs, dict) or set(inputs) != set(INPUT_BINDINGS):
        raise AuditError(f"{label}: input identity universe drift")
    if record.get("inputs_after") != inputs:
        raise AuditError(f"{label}: inputs changed during product execution")
    result: dict[str, dict[str, object]] = {}
    for name, expected in INPUT_BINDINGS.items():
        declared = inputs[name]
        if not isinstance(declared, dict) or set(declared) != {
            "path", "bytes", "sha256",
        }:
            raise AuditError(f"{label}/{name}: input identity shape drift")
        path = Path(str(declared["path"])).resolve()
        observed = base._identity(path)
        normalized = dict(declared)
        normalized["path"] = str(path)
        if observed != normalized or {
            key: observed[key] for key in ("bytes", "sha256")
        } != expected:
            raise AuditError(f"{label}/{name}: frozen input identity drift")
        result[name] = observed
    return result


def _normalize_loaded(value: object) -> object:
    if type(value) is bool:
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def _v(value: object) -> object:
    value = _normalize_loaded(value)
    if isinstance(value, str):
        return _CONTROL_WS.sub(" ", value).strip()
    return value


def _county(value: object) -> str:
    return ("" if value is None else str(value).strip().rstrip(".")).upper()


def _description(value: object) -> object:
    value = _v(value)
    if isinstance(value, str):
        return re.sub(r"\s+", " ", _DESC_PREFIX.sub("", value)).strip()
    return value


def _persisted(value: object) -> object:
    # openpyxl/OOXML cannot preserve an empty text value as a distinguishable
    # cell value when the output is reopened.  This is the sole blank fold.
    return None if value == "" else value


def _project_current(row: Sequence[object]) -> tuple[object, ...]:
    values = tuple(row) + (None,) * max(0, 10 - len(row))
    pm = "".join(
        "" if values[index] is None else str(values[index]).strip()
        for index in (3, 4, 5)
    )
    projected = (
        _v(values[0]), _county(values[1]), pm, _v(values[2]),
        _v(values[6]), _v(values[7]), _v(values[8]),
        _description(values[9]),
    )
    return tuple(_persisted(value) for value in projected)


def _project_tsn(row: Sequence[object]) -> tuple[object, ...]:
    values = list(tuple(row) + (None,) * max(0, 8 - len(row)))[:8]
    projected = [_v(value) for value in values]
    projected[1] = _county(projected[1])
    projected[7] = _description(projected[7])
    return tuple(_persisted(value) for value in projected)


def _ordered_digest(rows: Iterable[Sequence[object]]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update(json.dumps(
            list(row), ensure_ascii=False, separators=(",", ":"),
        ).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def _read_input_workbook(
    path: Path, *, raw_tsn: bool, expected_rows: int,
) -> tuple[list[tuple[object, ...]], list[tuple[object, ...]], dict[str, object]]:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        duplicates = sorted(name for name, count in Counter(names).items() if count > 1)
        bad_member = archive.testzip()
    if duplicates or bad_member is not None:
        raise AuditError(f"{path.name}: ZIP integrity/member uniqueness failed")

    sheet_name = "Highway Locations (TSN)" if raw_tsn else "Highway Locations"
    header = HEADERS if raw_tsn else CURRENT_HEADERS
    width = len(header)
    workbook = load_workbook(
        path, read_only=True, data_only=False, keep_links=False,
    )
    try:
        if workbook.sheetnames != [sheet_name]:
            raise AuditError(f"{path.name}: source sheet universe/order drift")
        sheet = workbook[sheet_name]
        if sheet.sheet_state != "visible":
            raise AuditError(f"{path.name}: source sheet is not visible")
        physical = iter(sheet.iter_rows())
        header_cells = tuple(next(physical, ()))
        observed_header = tuple(cell.value for cell in header_cells)
        if observed_header != header or any(
            cell.data_type in {"f", "e"} for cell in header_cells
        ):
            raise AuditError(f"{path.name}: exact source header drift")
        raw_rows: list[tuple[object, ...]] = []
        projected: list[tuple[object, ...]] = []
        padded = 0
        for physical_row, cells in enumerate(physical, 2):
            cells = tuple(cells)
            if len(cells) > width:
                raise AuditError(f"{path.name}: extra cell at row {physical_row}")
            if any(cell.data_type in {"f", "e"} for cell in cells):
                raise AuditError(f"{path.name}: formula/error at row {physical_row}")
            values = tuple(cell.value for cell in cells)
            if not any(value is not None for value in values):
                raise AuditError(f"{path.name}: blank physical row {physical_row}")
            if len(values) < width:
                padded += width - len(values)
                values += (None,) * (width - len(values))
            raw_rows.append(values)
            projected.append(
                _project_tsn(values) if raw_tsn else _project_current(values)
            )
    finally:
        workbook.close()
    if len(raw_rows) != expected_rows:
        raise AuditError(
            f"{path.name}: {len(raw_rows):,} source rows != {expected_rows:,}"
        )
    facts: dict[str, object] = {
        "identity": base._identity(path),
        "sheet": sheet_name, "headers": list(header),
        "rows": len(raw_rows), "columns": width,
        "padded_omitted_trailing_blanks": padded,
        "raw_rows_ordered_sha256": _ordered_digest(raw_rows),
        "projected_rows_ordered_sha256": _ordered_digest(projected),
        "zip_integrity_and_member_uniqueness": True,
        "no_formulas_errors_extra_cells_or_blank_rows": True,
    }
    if raw_tsn:
        pointers = Counter(row[6] for row in raw_rows)
        blank_county = sum(row[1] is None for row in raw_rows)
        if (
            blank_county != 46 or pointers["*P*"] != 283
            or pointers["-------->"] != 282
        ):
            raise AuditError("raw TSN source claim census drift")
        facts.update({
            "blank_county_rows": blank_county,
            "pointer_P": pointers["*P*"],
            "pointer_arrow": pointers["-------->"],
            "pointer_total": pointers["*P*"] + pointers["-------->"],
        })
    return raw_rows, projected, facts


def _validate_raw_twin_bundle(
    inputs: dict[str, dict[str, object]], raw_rows: list[tuple[object, ...]],
) -> dict[str, object]:
    result, result_identity = _strict_canonical_json(
        Path(inputs["raw_tsn_result"]["path"]), maximum=8_192,
    )
    manifest, manifest_identity = _strict_canonical_json(
        Path(inputs["raw_tsn_manifest"]["path"]), maximum=16_384,
    )
    provenance, provenance_identity = _strict_canonical_json(
        Path(inputs["raw_tsn_provenance"]["path"]), maximum=32_000_000,
    )

    if (
        result.get("audit") != "Highway Sequence raw-TSN development twin builder"
        or result.get("status") != "PASS" or result.get("terminal") is not True
        or result.get("not_an_acceptance_artifact") is not True
        or result.get("inputs_unchanged") is not True
        or result.get("counts") != RAW_COUNTS
    ):
        raise AuditError("raw-twin terminal result contract drift")
    invariants = result.get("invariants")
    if not isinstance(invariants, dict) or not invariants or not all(
        value is True for value in invariants.values()
    ):
        raise AuditError("raw-twin builder invariant is absent/red")
    artifacts = result.get("artifacts")
    if not isinstance(artifacts, dict):
        raise AuditError("raw-twin result artifact map absent")
    for artifact, input_name in {
        "workbook": "raw_tsn_workbook", "provenance": "raw_tsn_provenance",
        "manifest": "raw_tsn_manifest",
    }.items():
        if not _embedded_identity_matches(artifacts.get(artifact), inputs[input_name]):
            raise AuditError(f"raw-twin result {artifact} identity drift")

    if (
        manifest.get("audit") != "Highway Sequence raw-TSN development twin manifest"
        or manifest.get("not_an_acceptance_artifact") is not True
        or manifest.get("inputs_unchanged") is not True
        or manifest.get("inputs_before") != manifest.get("inputs_after")
    ):
        raise AuditError("raw-twin manifest terminal/input contract drift")
    cache = manifest.get("cache_validation")
    reopen = manifest.get("workbook_reopen")
    if not isinstance(cache, dict) or not isinstance(reopen, dict):
        raise AuditError("raw-twin cache/reopen evidence absent")
    raw_to_normalized = cache.get("raw_to_normalized")
    expected_loss = {
        "normalized_rows": 69_758,
        "projected_rows": 69_758,
        "typed_cell_delta_count": 566,
        "typed_cell_deltas_by_field": {
            "Description": 1, "Distance To Next Point": 565,
        },
        "pointer_delta_raw_tokens": {"*P*": 283, "-------->": 282},
    }
    if (
        cache.get("counts") != RAW_COUNTS
        or not isinstance(raw_to_normalized, dict)
        or any(raw_to_normalized.get(key) != value for key, value in expected_loss.items())
        or reopen.get("headers") != list(HEADERS)
        or reopen.get("data_rows") != 69_804
        or reopen.get("blank_county_rows") != 46
        or reopen.get("pointer_P") != 283
        or reopen.get("pointer_arrow") != 282
        or reopen.get("typed_rows_exact") is not True
    ):
        raise AuditError("raw-twin conservation/reopen evidence drift")
    raw_digest = _ordered_digest(raw_rows)
    if (
        raw_digest != result.get("ordered_rows_sha256")
        or raw_digest != reopen.get("ordered_rows_sha256")
        or raw_digest != cache.get("twin_rows_ordered_sha256")
    ):
        raise AuditError("raw-twin workbook ordered-row digest disagreement")

    before = manifest["inputs_before"]
    if not isinstance(before, dict):
        raise AuditError("raw-twin source binding tree absent")
    authenticated_upstream = {
        label: _authenticate_embedded_file(value, f"raw twin upstream/{label}")
        for label, value in before.items()
    }
    embedded_accepted = manifest.get("embedded_accepted_bindings")
    if not isinstance(embedded_accepted, dict):
        raise AuditError("raw-twin embedded acceptance bindings absent")
    for label, value in embedded_accepted.items():
        identity = authenticated_upstream.get(label)
        if identity is None or not _embedded_identity_matches(value, identity):
            raise AuditError(f"raw-twin embedded acceptance {label} drift")

    expected_provenance_fields = {
        "audit", "development_cache_binding", "embedded_accepted_bindings",
        "not_an_acceptance_artifact", "raw_documents", "raw_members",
        "reason", "row_count", "rows", "schema",
    }
    if (
        set(provenance) != expected_provenance_fields
        or provenance.get("audit")
        != "Highway Sequence raw-TSN development twin provenance"
        or provenance.get("not_an_acceptance_artifact") is not True
        or provenance.get("schema") != {
            "sheet": "Highway Locations (TSN)", "headers": list(HEADERS),
        }
        or provenance.get("row_count") != 69_804
    ):
        raise AuditError("raw-twin provenance envelope/schema drift")
    if not _embedded_identity_matches(
        provenance.get("development_cache_binding"),
        authenticated_upstream["development_row_cache"],
    ) or provenance.get("embedded_accepted_bindings") != embedded_accepted:
        raise AuditError("raw-twin provenance binding chain drift")

    raw_members = provenance.get("raw_members")
    raw_documents = provenance.get("raw_documents")
    rows = provenance.get("rows")
    if not isinstance(raw_members, list) or not isinstance(raw_documents, list) or not isinstance(rows, list):
        raise AuditError("raw-twin provenance arrays absent")
    expected_names = [f"D{district:02d} HSL TSN.pdf" for district in range(1, 13)]
    member_names = [item.get("name") for item in raw_members if isinstance(item, dict)]
    if member_names != expected_names or len(raw_documents) != 12 or len(rows) != 69_804:
        raise AuditError("raw-twin provenance member/document/row census drift")
    if any(
        not isinstance(item, dict) or set(item) != {"name", "bytes", "sha256"}
        or type(item["bytes"]) is not int or item["bytes"] <= 0
        or not isinstance(item["sha256"], str) or len(item["sha256"]) != 64
        for item in raw_members
    ):
        raise AuditError("raw-twin provenance raw-member identity shape drift")

    kinds: Counter[str] = Counter()
    blank_county_kinds: Counter[str] = Counter()
    provenance_digest = hashlib.sha256()
    for ordinal, (item, workbook_values) in enumerate(zip(rows, raw_rows, strict=True)):
        if not isinstance(item, dict) or set(item) != {
            "workbook_row", "source_ref", "source_context", "workbook_values",
        } or item["workbook_row"] != ordinal + 2:
            raise AuditError(f"raw-twin provenance row {ordinal}: shape/ordinal drift")
        if item["workbook_values"] != list(workbook_values):
            raise AuditError(f"raw-twin provenance row {ordinal}: workbook value drift")
        source_ref = item["source_ref"]
        context = item["source_context"]
        if (
            not isinstance(source_ref, dict) or set(source_ref) != {
                "member", "physical_page", "printed_page", "line", "top",
            }
            or source_ref["member"] not in expected_names
            or type(source_ref["physical_page"]) is not int
            or type(source_ref["printed_page"]) is not int
            or type(source_ref["line"]) is not int
            or not isinstance(source_ref["top"], str)
            or not isinstance(context, dict) or set(context) != {
                "district", "direction", "record_kind", "raw_text",
            }
            or context["record_kind"] not in {"data", "equate"}
            or not isinstance(context["raw_text"], str) or not context["raw_text"]
        ):
            raise AuditError(f"raw-twin provenance row {ordinal}: source claim drift")
        kinds[context["record_kind"]] += 1
        if workbook_values[1] is None:
            blank_county_kinds[context["record_kind"]] += 1
        provenance_digest.update(base._json_line({
            "source_ref": source_ref, "source_context": context,
            "workbook_values": list(workbook_values),
        }))
    if kinds != Counter({"data": 68_806, "equate": 998}) or blank_county_kinds != Counter({"equate": 46}):
        raise AuditError("raw-twin provenance record-kind conservation drift")

    return {
        "result": result_identity, "manifest": manifest_identity,
        "provenance": provenance_identity,
        "status": "PASS", "terminal": True,
        "not_an_acceptance_artifact": True,
        "counts": RAW_COUNTS,
        "ordered_rows_sha256": raw_digest,
        "provenance_claim_rows_sha256": provenance_digest.hexdigest(),
        "raw_members": 12, "raw_documents": 12,
        "authenticated_upstream_bindings": authenticated_upstream,
        "normalization_losses_bound": {
            "dropped_pre_county_equates": 46,
            "blanked_pointer_tokens": 565,
            "altered_description_cells": 1,
            "typed_cell_deltas": 566,
        },
        "provenance_rows_exact_to_workbook": True,
    }


def _validate_preimport(
    record: dict[str, object], inputs: dict[str, dict[str, object]], label: str,
) -> dict[str, object]:
    value = record.get("raw_twin_preimport_validation")
    if not isinstance(value, dict):
        raise AuditError(f"{label}: pre-import raw-twin validation absent")
    if (
        value.get("validated_before_product_import") is not True
        or value.get("terminal_status") != "PASS"
        or value.get("terminal") is not True
        or value.get("not_an_acceptance_artifact") is not True
        or value.get("counts") != RAW_COUNTS
        or value.get("headers") != list(HEADERS)
    ):
        raise AuditError(f"{label}: pre-import raw-twin validation drift")
    identities = value.get("identity_bound_artifacts")
    expected = {
        name: inputs[name] for name in (
            "raw_tsn_workbook", "raw_tsn_provenance",
            "raw_tsn_manifest", "raw_tsn_result",
        )
    }
    if identities != expected:
        raise AuditError(f"{label}: pre-import raw identity map drift")
    return {
        "validated_before_product_import": True,
        "terminal_status": "PASS", "counts": RAW_COUNTS,
        "identity_bound_artifacts_exact": True,
    }


def _load_leg_result(path: Path, label: str) -> tuple[dict[str, object], dict[str, object]]:
    record, identity = _strict_canonical_json(path, maximum=2_000_000)
    if (
        record.get("audit")
        != "Stage 8 Highway Sequence development raw-twin product comparison leg"
        or record.get("leg") != label
        or record.get("not_an_acceptance_artifact") is not True
        or Path(str(record.get("output_root", ""))).resolve() != path.parent.resolve()
    ):
        raise AuditError(f"{label}: raw product witness envelope drift")
    reason = record.get("reason")
    if not isinstance(reason, str) or "cache-derived raw-TSN twin" not in reason:
        raise AuditError(f"{label}: development-only warning absent")
    invariants = record.get("invariants")
    if not isinstance(invariants, dict) or not invariants or not all(
        value is True for value in invariants.values()
    ):
        raise AuditError(f"{label}: runner invariant is absent/red")
    actual_names = sorted(item.name for item in path.parent.iterdir())
    if record.get("expected_final_artifact_names") != actual_names:
        raise AuditError(f"{label}: final artifact universe drift")
    return record, identity


def _compare_source_rows(
    actual: list[base.SourceRow], expected: list[tuple[object, ...]], label: str,
) -> dict[str, object]:
    if len(actual) != len(expected):
        raise AuditError(f"{label}: embedded/external row census drift")
    digest = hashlib.sha256()
    for index, (row, expected_values) in enumerate(zip(actual, expected, strict=True)):
        if row.projection != expected_values:
            raise AuditError(
                f"{label}: embedded source differs from input projection at "
                f"source row {index + 2}: {row.projection!r} != {expected_values!r}"
            )
        digest.update(base._json_line([index, *expected_values]))
    return {
        "rows": len(actual), "cells": len(actual) * 8,
        "ordered_projection_sha256": digest.hexdigest(),
        "every_embedded_source_row_and_cell_exact_to_input": True,
    }


def _validate_leg_artifact_tree(
    label: str, path: Path, record: dict[str, object],
) -> dict[str, object]:
    root = path.parent.resolve()
    artifact_path, artifact_identity = base._declared_identity(
        record.get("artifact_manifest"), f"{label} artifact manifest",
    )
    product_path, product_identity = base._declared_identity(
        record.get("product_code_manifest"), f"{label} product manifest",
    )
    if artifact_path != root / "artifact-manifest.json" or product_path != root / "product-code-manifest.json":
        raise AuditError(f"{label}: audit manifest path drift")
    if artifact_path.read_bytes() != base._json_line(record.get("artifact_manifest_before_result")):
        raise AuditError(f"{label}: artifact manifest content drift")
    if product_path.read_bytes() != base._json_line(record.get("loaded_product_code")):
        raise AuditError(f"{label}: product manifest content drift")
    tree = base._validate_tree_manifest(
        root, record.get("artifact_manifest_before_result"),
        ignored_names=frozenset(("artifact-manifest.json", "result.json")),
        label=f"{label} artifact tree",
    )
    residue = record.get("residue_gate")
    if not isinstance(residue, dict) or residue.get("transient_residue") != []:
        raise AuditError(f"{label}: terminal residue gate drift")
    return {
        "tree": tree, "result": base._identity(path),
        "artifact_manifest": artifact_identity,
        "product_code_manifest": product_identity,
        "exact_final_artifact_universe": True,
        "no_transient_residue": True,
    }


def _parse_leg_arguments(values: list[str]) -> dict[str, Path]:
    if len(values) != 2:
        raise AuditError("--leg-result must be repeated exactly twice")
    result: dict[str, Path] = {}
    for value in values:
        if "=" not in value:
            raise AuditError(f"--leg-result must be LEG=PATH, got {value!r}")
        label, raw_path = value.split("=", 1)
        if label not in LEG_SPECS or label in result or not raw_path:
            raise AuditError(f"invalid/duplicate --leg-result label: {label!r}")
        path = Path(raw_path).expanduser().resolve()
        if path.name != "result.json":
            raise AuditError(f"{label}: --leg-result must name result.json")
        result[label] = path
    if set(result) != set(LEG_SPECS):
        raise AuditError("--leg-result must name both raw-TSN product legs")
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--leg-result", action="append", required=True, metavar="LEG=PATH",
        help="repeat exactly twice: excel_vs_raw_tsn and pdf_vs_raw_tsn",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args(argv)
    leg_paths = _parse_leg_arguments(args.leg_result)

    records: dict[str, dict[str, object]] = {}
    result_identities: dict[str, dict[str, object]] = {}
    input_maps: dict[str, dict[str, dict[str, object]]] = {}
    for label, path in leg_paths.items():
        records[label], result_identities[label] = _load_leg_result(path, label)
        input_maps[label] = _bind_inputs(records[label], label)
    inputs = input_maps["excel_vs_raw_tsn"]
    if input_maps["pdf_vs_raw_tsn"] != inputs:
        raise AuditError("the two raw product legs disagree on frozen inputs")

    excel_raw, excel_projected, excel_source = _read_input_workbook(
        Path(inputs["current_excel"]["path"]), raw_tsn=False,
        expected_rows=60_494,
    )
    pdf_raw, pdf_projected, pdf_source = _read_input_workbook(
        Path(inputs["current_pdf"]["path"]), raw_tsn=False,
        expected_rows=60_493,
    )
    raw_rows, tsn_projected, raw_source = _read_input_workbook(
        Path(inputs["raw_tsn_workbook"]["path"]), raw_tsn=True,
        expected_rows=69_804,
    )
    raw_bundle = _validate_raw_twin_bundle(inputs, raw_rows)
    external = {
        "current_excel": (excel_raw, excel_projected, excel_source),
        "current_pdf": (pdf_raw, pdf_projected, pdf_source),
        "raw_tsn": (raw_rows, tsn_projected, raw_source),
    }

    audited: dict[str, object] = {}
    generation_ids: set[str] = set()
    referenced_payloads: set[Path] = set()
    declared_payloads: set[Path] = set()
    product_code = {}
    trees = {}
    preimport = {}
    for label, spec in LEG_SPECS.items():
        record = records[label]
        path = leg_paths[label]
        trees[label] = _validate_leg_artifact_tree(label, path, record)
        product_code[label] = base._validate_loaded_product_code(
            record.get("loaded_product_code")
        )
        preimport[label] = _validate_preimport(record, inputs, label)
        formulas_path, values_path, output_identities = base._bind_output_paths(
            label, record, spec,
        )
        workbook = base._inspect_leg_workbooks(
            label, spec, formulas_path, values_path,
        )
        side_a_expected = external[str(spec["side_a_input"])][1]
        exact_a = _compare_source_rows(
            workbook["rows_a"], side_a_expected, f"{label} side A",
        )
        exact_b = _compare_source_rows(
            workbook["rows_b"], tsn_projected, f"{label} raw TSN",
        )
        publication, referenced = base._inspect_publication(
            label, formulas_path, values_path, record,
            workbook["comparison"]["counts"], workbook["rows_a"],
            workbook["rows_b"],
            workbook["comparison"]["paired_source_indices"], output_identities,
        )
        generation_id = publication["generation_id"]
        if generation_id in generation_ids:
            raise AuditError("raw product legs reused one generation ID")
        generation_ids.add(generation_id)
        referenced_payloads.update(referenced)
        declared_payloads.update(
            member.resolve() for member in path.parent.glob("*.comparison-payload.zlib")
            if member.is_file()
        )
        counts = workbook["comparison"]["counts"]
        expected_shape = {
            key: int(spec[key]) for key in (
                "paired_rows", "side_a_only_rows", "side_b_only_rows",
            )
        }
        if {key: counts[key] for key in expected_shape} != expected_shape:
            raise AuditError(f"{label}: complete raw-source shape drift")
        public_workbook = base._public_leg_result(workbook)
        public_workbook["external_source_projection"] = {
            str(spec["side_a"]): exact_a, "TSN": exact_b,
        }
        audited[label] = {
            "result": result_identities[label],
            "outputs": output_identities,
            "workbook": public_workbook,
            "publication": publication,
            "complete_raw_source_shape": expected_shape,
        }
    if referenced_payloads != declared_payloads:
        raise AuditError("raw product payload universe has a missing/orphan chunk")

    result = {
        "audit": "Stage 8 Highway Sequence product comparisons against raw-TSN development twin",
        "status": "pass_with_product_defects",
        "acceptance_artifact": False,
        "reason_not_acceptance": (
            "The raw TSN workbook is a cache-derived development twin, so final "
            "acceptance must reparse the immutable TSN PDFs. This artifact faithfully "
            "authenticates current product output while preserving evidence that the "
            "accepted normalized TSN workbook drops 46 pre-county equates, blanks 565 "
            "pointer tokens, and alters one Description cell."
        ),
        "leg_results": result_identities,
        "inputs": inputs,
        "input_workbooks": {
            "current_excel": excel_source,
            "current_pdf": pdf_source,
            "raw_tsn": raw_source,
        },
        "raw_tsn_bundle": raw_bundle,
        "runner_preimport_validation": preimport,
        "loaded_product_code": product_code,
        "artifact_trees": trees,
        "legs": audited,
        "known_normalization_claim_losses": {
            "dropped_rows": {"pre_county_equates": 46},
            "blanked_cells": {
                "Distance To Next Point *P*": 283,
                "Distance To Next Point -------->": 282,
            },
            "altered_cells": {"Description punctuation": 1},
            "total_typed_cell_deltas": 566,
        },
        "verified_invariants": {
            "exactly_two_terminal_leg_results": len(records) == 2,
            "all_six_input_files_authenticated_and_unchanged": all(
                item == inputs for item in input_maps.values()
            ),
            "raw_tsn_has_69804_rows": raw_source["rows"] == 69_804,
            "raw_tsn_has_46_blank_county_rows": raw_source["blank_county_rows"] == 46,
            "raw_tsn_has_565_pointer_tokens": raw_source["pointer_total"] == 565,
            "raw_tsn_provenance_exact_to_all_rows": raw_bundle["provenance_rows_exact_to_workbook"],
            "both_formula_value_twins_streamed_and_reconstructed": all(
                item["workbook"]["comparison"]["comparison_rows_reconstructed_from_sources"]
                for item in audited.values()
            ),
            "all_embedded_source_rows_cells_exact_to_bound_inputs": all(
                all(source["every_embedded_source_row_and_cell_exact_to_input"]
                    for source in item["workbook"]["external_source_projection"].values())
                for item in audited.values()
            ),
            "excel_complete_raw_shape_57072_3422_12732": audited["excel_vs_raw_tsn"]["complete_raw_source_shape"] == {
                "paired_rows": 57_072, "side_a_only_rows": 3_422,
                "side_b_only_rows": 12_732,
            },
            "pdf_complete_raw_shape_57505_2988_12299": audited["pdf_vs_raw_tsn"]["complete_raw_source_shape"] == {
                "paired_rows": 57_505, "side_a_only_rows": 2_988,
                "side_b_only_rows": 12_299,
            },
            "all_returned_persisted_workbook_counts_exact": all(
                item["publication"]["returned_persisted_workbook_counts_exact"]
                for item in audited.values()
            ),
            "all_pairing_traces_independently_exact": all(
                item["publication"]["pairing_trace"]["all_assignments_independently_exact_and_lexicographic"]
                for item in audited.values()
            ),
            "all_sidecars_chunks_generation_and_tree_manifests_exact": referenced_payloads == declared_payloads,
            "no_differing_cell_count_was_assumed": True,
            "development_non_acceptance_boundary_explicit": True,
        },
    }
    if not all(result["verified_invariants"].values()):
        raise AuditError("final raw product comparison invariant is red")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_bytes(base._json_line(result))
    print(f"PASS raw-TSN Highway Sequence product comparison audit: {args.output}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (AuditError, base.AuditError) as exc:
        print(f"FAIL raw-TSN Highway Sequence product comparison audit: {exc}")
        raise SystemExit(1)
