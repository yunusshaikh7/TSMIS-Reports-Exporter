"""Golden check for CMP-AUD-032 — flat cross-env schema pinning.

Before: the flat families (Ramp Detail, Highway Sequence, Highway Detail,
Intersection Detail) carried no canonical header, so `_load_xlsx_side` locked the
header from the alphabetically-first readable file. Two identically-malformed or
legacy sides then compared as a clean MATCH, a stray same-sheet workbook could
skip valid files or flip the result, and Highway Detail never called its exact
recognizer. (Highway Log was already pinned via `_hl_canonical_header`.)

After: each flat family pins its EXACT current export schema through a
`header_canonicalizer` (the `_flat_header_recognizer` factory, or ID's
current/legacy `_id_canonical_header`): a recognized layout canonicalizes;
anything else returns None → the comparison refuses with "not a recognized
column layout". So two identically-malformed / legacy / truncated / reordered
sides fail instead of pairing on a trusted-first-file header.

The canonical headers are the vs-TSN comparators' `_TSMIS_HEADER[1:]`
(Ramp Detail / Highway Sequence / Intersection Detail) and
`highway_detail_columns.HEADER` — verified against the real 7.9 statewide
exports (126 RD / 252 HSL / 252 HD / 217 ID files, each ONE consistent header
shape, none carrying a leading Route column).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_env_flat_schema.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
from events import Events
from openpyxl import Workbook

_fail = []


def check(name, ok, extra=""):
    print(f"  [{'OK ' if ok else 'FAIL'}] {name}" + (f"  {extra}" if extra else ""))
    if not ok:
        _fail.append(name)


def _relabel(raw):
    """The on-disk export header as `_load_xlsx_side` presents it (blank/None
    internal cells become the positional '(col X)' label)."""
    from openpyxl.utils import get_column_letter
    return [str(c).strip() if (c is not None and str(c).strip())
            else f"(col {get_column_letter(i + 1)})" for i, c in enumerate(raw)]


def _families():
    """(family, recognizer, raw_current_header_without_route)."""
    import compare_ramp_detail_tsn as rd
    import compare_highway_sequence_tsn as hsl
    import compare_intersection_detail_tsn as idt
    import highway_detail_columns as hdc
    return [
        ("Ramp Detail", compare_env._ramp_detail_canonical_header,
         list(rd._TSMIS_HEADER[1:])),
        ("Highway Sequence", compare_env._highway_sequence_canonical_header,
         list(hsl._TSMIS_HEADER[1:])),
        ("Highway Detail", compare_env._highway_detail_canonical_header,
         list(hdc.HEADER)),
        ("Intersection Detail", compare_env._id_canonical_header,
         list(idt._TSMIS_HEADER[1:])),
    ]


def test_recognizers():
    print("Recognizer matrix (current recognized; malformed/legacy/truncated/reordered refused):")
    for name, canon, raw in _families():
        loaded = _relabel(raw)               # what the loader hands the canonicalizer
        got = canon(list(loaded))
        check(f"{name}: current export layout is recognized",
              got is not None and got == loaded)
        check(f"{name}: a bogus 2-column layout is refused",
              canon(["PM", "Bogus"]) is None)
        check(f"{name}: a truncated layout (one column dropped) is refused",
              canon(loaded[:-1]) is None)
        if len(loaded) >= 3:
            reordered = [loaded[1], loaded[0]] + loaded[2:]
            check(f"{name}: a reordered layout is refused",
                  reordered == loaded or canon(reordered) is None)


def _write_flat(folder, sheet, route, header, row):
    folder.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(header))
    ws.append(list(row))
    # any name matching the ..._route_<n>.xlsx contract (CMP-AUD-031)
    wb.save(folder / f"x_route_{route}.xlsx")
    wb.close()


def _rd_real_row():
    import compare_ramp_detail_tsn as rd
    hdr = list(rd._TSMIS_HEADER[1:])
    row = [""] * len(hdr)
    row[0] = "01-DN-101"
    row[hdr.index("PM")] = "1.000"
    row[hdr.index("Description")] = "RAMP A"
    return hdr, row


def test_end_to_end_bogus_refused():
    print("End-to-end: two identically-BOGUS Ramp Detail sides refuse (no phantom match):")
    root = Path(tempfile.mkdtemp(prefix="rd032_"))
    a = root / "2026-06-19 ssor-prod" / "ramp_detail"
    b = root / "2026-06-19 ars-prod" / "ramp_detail"
    for side in (a, b):
        _write_flat(side, "TSAR - Ramp Detail", "101", ["PM", "Bogus"], ["1.000", "x"])
    res = compare_env.RAMP_DETAIL.compare_folders(
        str(a.parent), str(b.parent), str(root / "cmp.xlsx"), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    check("two bogus sides are REFUSED (status error)", res.status == "error")
    check("the message names the unrecognized layout",
          "recognized" in (res.message or "").lower(), res.message)
    check("no comparison workbook was written", not (root / "cmp.xlsx").exists())


def test_end_to_end_current_ok():
    print("End-to-end: two real current Ramp Detail sides compare cleanly:")
    root = Path(tempfile.mkdtemp(prefix="rd032ok_"))
    a = root / "2026-06-19 ssor-prod" / "ramp_detail"
    b = root / "2026-06-19 ars-prod" / "ramp_detail"
    hdr, row = _rd_real_row()
    for side in (a, b):
        _write_flat(side, "TSAR - Ramp Detail", "101", hdr, row)
    res = compare_env.RAMP_DETAIL.compare_folders(
        str(a.parent), str(b.parent), str(root / "cmp.xlsx"), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    check("two real-layout sides compare (status ok)", res.status == "ok", res.message)
    check("the comparison workbook was written", (root / "cmp.xlsx").exists())


def test_end_to_end_current_vs_bogus_refused():
    print("End-to-end: a real side vs a bogus side refuses:")
    root = Path(tempfile.mkdtemp(prefix="rd032mix_"))
    a = root / "2026-06-19 ssor-prod" / "ramp_detail"
    b = root / "2026-06-19 ars-prod" / "ramp_detail"
    hdr, row = _rd_real_row()
    _write_flat(a, "TSAR - Ramp Detail", "101", hdr, row)
    _write_flat(b, "TSAR - Ramp Detail", "101", ["PM", "Bogus"], ["1.000", "x"])
    res = compare_env.RAMP_DETAIL.compare_folders(
        str(a.parent), str(b.parent), str(root / "cmp.xlsx"), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    check("real-vs-bogus is REFUSED (status error)", res.status == "error", res.message)


def main():
    test_recognizers()
    test_end_to_end_bogus_refused()
    test_end_to_end_current_ok()
    test_end_to_end_current_vs_bogus_refused()
    if _fail:
        print(f"\nFAILED: {len(_fail)}")
        raise SystemExit(1)
    print("\nALL COMPARE-ENV-FLAT-SCHEMA (CMP-AUD-032) CHECKS PASSED")


if __name__ == "__main__":
    main()
