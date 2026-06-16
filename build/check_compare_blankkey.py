"""Golden check for COMPARE-BLANK-KEYFIELD-SELFCHECK (scripts/compare_core.py).

Reproduces the blank-key-field bug that made the cross-environment Highway
Sequence comparison fail 6 of 9 SELF-CHECK rows AND mis-compare every blank-key
row. Run with the build venv:

    build\\.venv\\Scripts\\python.exe build\\check_compare_blankkey.py

Excel-COM-verified before/after on this exact fixture (2026-06-15):
    OLD code: 6/9 SELF-CHECK = CHECK; verdict WRONG ("1 differing, 4 one-sided")
    FIXED   : 9/9 SELF-CHECK = OK;    verdict correct ("2 differing, 2 one-sided")

The live-formula SELF-CHECK reconciliation needs Excel to evaluate, so this
script guards the fix WITHOUT Excel by asserting the structural changes the
self-checks depend on (literal lookup keys instead of a live COUNTIFS that
mis-numbers blank keys; row counts taken off an always-present column instead
of the blank-able key field) plus the engine's own pure-Python diff counts.
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from compare_core import (CompareSchema, count_diffs, keys_for, run_compare,
                          union_keys)
from openpyxl import load_workbook

SC = CompareSchema(report_name="BlankKeyTest", header=["County", "PM", "Desc"],
                   side_a="AENV", side_b="BENV", id_noun="row",
                   id_noun_plural="rows", sides_noun="environments")
# [Route, County(key — BLANK on two rows), PM, Desc]
ROWS_A = [["001", "ORA", "0.1", "d1"], ["001", "ORA", "0.2", "d2"],
          ["001", "", "0.3", "b1"], ["001", "", "0.4", "b2"],
          ["001", "LA", "0.5", "onlyA"]]
ROWS_B = [["001", "ORA", "0.1", "d1"], ["001", "ORA", "0.2", "d2X"],
          ["001", "", "0.3", "b1"], ["001", "", "0.4", "b2X"],
          ["001", "SD", "0.6", "onlyB"]]


def main():
    # 1) Data correctness: blank-key rows must align as Both (not dropped), and
    #    BOTH Desc diffs (incl. the blank-key one) must be counted.
    kt, kn = keys_for(ROWS_A, True), keys_for(ROWS_B, True)
    u = union_keys(kt, kn)
    c = count_diffs(SC, ROWS_A, ROWS_B, kt, kn, u, True)
    assert c["both"] == 4, ("both should be 4", c["both"])
    assert c["t_only"] == 1 and c["n_only"] == 1, ("one-sided", c)
    assert c["diff_cells"] == 2, ("diff_cells should be 2", c["diff_cells"])

    # 2) Structural: the live-formulas workbook must carry the fix.
    out = os.path.join(tempfile.gettempdir(), "_blankkey_check.xlsx")
    res = run_compare(SC, ROWS_A, ROWS_B, True, out, mode="formulas")
    assert res.status == "ok", res.status
    wb = load_workbook(out, data_only=False)

    # 2a) Bug B: the data-sheet "Key (helper)" column is a LITERAL string, not a
    #     live COUNTIFS (which mis-numbers blank keys via Excel's blank-criterion
    #     quirk). Row 2 of side A is 001/ORA/occurrence 1.
    kcell = wb["AENV"].cell(row=2, column=wb["AENV"].max_column)
    assert kcell.data_type != "f" and "COUNTIF" not in str(kcell.value), \
        ("Key (helper) must be a literal, not COUNTIFS", kcell.value)
    assert str(kcell.value) == "001|ORA|1", ("key value", kcell.value)

    # 2b) Bug A: the row-count self-checks count an always-present column (the
    #     back-link A / the numeric occurrence #), never the blank-able key
    #     field (County = column C on the data sheets here).
    formulas = [str(cl.value) for rw in wb["Summary"].iter_rows()
                for cl in rw if cl.data_type == "f"]
    blob = "\n".join(formulas)
    assert "COUNTA(AENV!C:C)" not in blob and "COUNTA(BENV!C:C)" not in blob, \
        "row counts still use the blank-able key-field column (COUNTA of County)"
    assert "COUNTA(AENV!A:A)" in blob, "data-row count not taken off back-link col A"
    assert any("COUNT(Comparison!" in f for f in formulas), \
        "union count not taken off the numeric occurrence column"
    os.remove(out)
    print("OK  COMPARE-BLANK-KEYFIELD-SELFCHECK: blank-key rows align "
          "(both=4, 2 diffs, 2 one-sided); keys literal; counts off "
          "back-link/occurrence.")


if __name__ == "__main__":
    main()
