"""Hermetic CMP-AUD-035 raw TSN source-admission contract.

Proves the production boundaries, without reading the development corpus:

* a single-statewide normalizer admits exactly one ordinary matching source;
  zero and two candidates fail, owner-lock files are ignored, and mtime never
  chooses a winner;
* Ramp/Intersection/Highway Detail require their exact complete ordered raw
  headers before either raw projector runs;
* each detail builder returns explicit COMPLETE producer state with zero
  skipped/failed inputs, while its emitted workbook independently contains the
  admitted row and exact normalized header;
* Ramp Summary and Intersection Summary inherit the same exact-one boundary.

Run from the repository root:
    build\\.venv\\Scripts\\python.exe -X utf8 build\\check_tsn_raw_source_contract.py
"""
import contextlib
import os
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import Workbook, load_workbook  # noqa: E402

import compare_highway_detail_tsn as hdt  # noqa: E402
import compare_intersection_detail_tsn as idt  # noqa: E402
import compare_intersection_summary_tsn as istsn  # noqa: E402
import compare_ramp_detail_tsn as rd  # noqa: E402
import compare_ramp_summary_tsn as rstsn  # noqa: E402
import artifact_store  # noqa: E402
from events import ConsolidateResult  # noqa: E402
import outcome  # noqa: E402
import tsn_library  # noqa: E402
import tsn_load_highway_detail as hd_load  # noqa: E402
import tsn_load_intersection_detail as id_load  # noqa: E402
import tsn_load_intersection_summary as is_load  # noqa: E402
import tsn_load_ramp_detail as rd_load  # noqa: E402
import tsn_load_ramp_summary as rs_load  # noqa: E402


_fail = []


def check(name, condition):
    print(f"  [{'OK ' if condition else 'FAIL'}] {name}")
    if not condition:
        _fail.append(name)


@contextlib.contextmanager
def patch(obj, name, value):
    old = getattr(obj, name)
    setattr(obj, name, value)
    try:
        yield
    finally:
        setattr(obj, name, old)


def write_raw(path, header, row=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.append(list(header))
    ws.append(list(row if row is not None else ["x"] + [None] * (len(header) - 1)))
    wb.save(path)
    wb.close()


def workbook_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return ws.title, [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def write_tsn_certificate(report, workbook, *, raw_manifest=None,
                          normalization_version=None, workbook_identity=None,
                          identity_token=None):
    """Publish the complete strict TSN provenance modeled by production."""
    spec = tsn_library.get(report)
    raw_manifest = (tsn_library._raw_manifest(report)
                    if raw_manifest is None else raw_manifest)
    workbook_identity = (
        tsn_library.normalized_workbook_identity(workbook)
        if workbook_identity is None else workbook_identity)
    if identity_token is None:
        try:
            identity_token = tsn_library.canonical_normalized_identity_token(
                report, raw_manifest, workbook_identity)
        except ValueError:
            identity_token = "invalid-provenance-cannot-have-a-valid-token"
    return tsn_library.consolidation_meta.write_outcome(
        workbook, ConsolidateResult(status="ok", completion=outcome.COMPLETE),
        extra={
            "tsn_normalization_version": (
                spec.normalization_version if normalization_version is None
                else normalization_version),
            "tsn_raw_manifest": raw_manifest,
            "tsn_normalized_workbook_identity": workbook_identity,
            "tsn_artifact_identity_token": identity_token,
        })


def test_shared_exact_one():
    print("shared statewide admission (0 / 1 / 2; owner-lock; no mtime winner):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_admission_"))
    try:
        raw = root / "raw"
        raw.mkdir()
        calls = []

        def project(path):
            calls.append(Path(path).name)

            def result(_out_name):
                return ConsolidateResult(status="ok", completion=outcome.COMPLETE)
            return [["emitted"]], result

        kwargs = dict(
            glob="*.xlsx", deps_ok=True, deps_msg="deps", no_raw_what="fixture .xlsx",
            no_raw_hint="add one", log_label="Fixture", sheet="Normalized",
            header=["Value"], header_align={"horizontal": "center"}, project=project)

        zero = tsn_library.build_normalized(raw, root / "zero.xlsx", **kwargs)
        check("zero candidates fails before projection",
              zero.status == "error" and "No raw fixture .xlsx" in zero.message and not calls)

        ordinary = raw / "truth.xlsx"
        ordinary.write_bytes(b"ordinary")
        (raw / "~$truth.xlsx").write_bytes(b"owner-lock")
        one_out = root / "one.xlsx"
        one = tsn_library.build_normalized(raw, one_out, **kwargs)
        title, rows = workbook_rows(one_out)
        check("one ordinary candidate admitted; owner-lock ignored",
              one.status == "ok" and calls == ["truth.xlsx"]
              and len(tsn_library._statewide_raw_candidates(raw, "*.xlsx")) == 1)
        check("independent workbook read sees exactly one emitted row",
              title == "Normalized" and rows == [["Value"], ["emitted"]])

        calls.clear()
        older = raw / "older.xlsx"
        older.write_bytes(b"older")
        os.utime(older, (1000.0, 1000.0))
        os.utime(ordinary, (9000.0, 9000.0))
        two_out = root / "two.xlsx"
        two = tsn_library.build_normalized(raw, two_out, **kwargs)
        check("two ordinary candidates fail; newest mtime is not selected",
              two.status == "error" and "Found 2 ordinary matching files" in two.message
              and "older.xlsx" in two.message and "truth.xlsx" in two.message
              and not calls and not two_out.exists())
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_statewide_immutable_snapshot_and_commit_guard():
    print("statewide builders parse captured bytes and preserve last-good on drift:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_snapshot_"))
    try:
        raw = root / "raw"
        raw.mkdir()
        member = raw / "truth.xlsx"
        member.write_bytes(b"A")
        out = root / "normalized.xlsx"

        def result(_out_name):
            return ConsolidateResult(status="ok", completion=outcome.COMPLETE)

        kwargs = dict(
            glob="*.xlsx", deps_ok=True, deps_msg="deps",
            no_raw_what="fixture .xlsx", no_raw_hint="add one",
            log_label="Fixture", sheet="Normalized", header=["Value"],
            header_align={"horizontal": "center"})

        observed = []

        def transient_project(path):
            # Mutate the live source before reading the parser input.  Only a
            # private captured snapshot can still yield A here.
            member.write_bytes(b"B")
            observed.append((Path(path).resolve() != member.resolve(),
                             Path(path).read_bytes()))
            member.write_bytes(b"A")
            return [["from-A"]], result

        transient = tsn_library.build_normalized(
            raw, out, project=transient_project, **kwargs)
        _title, transient_rows = workbook_rows(out)
        check("transient A->B->A mutation cannot alter the parsed generation",
              transient.status == "ok" and observed == [(True, b"A")]
              and transient_rows == [["Value"], ["from-A"]]
              and tsn_library._raw_contract.validate_raw_manifest(
                  transient.tsn_raw_manifest) == transient.tsn_raw_manifest)
        check("private source snapshot is removed after projection",
              not list(root.glob(".tsn-statewide-source-*")))

        def persistent_project(path):
            member.write_bytes(b"B")
            return [[Path(path).read_bytes().decode("ascii")]], result

        persistent = tsn_library.build_normalized(
            raw, out, project=persistent_project, **kwargs)
        _title, after_persistent = workbook_rows(out)
        check("persistent mutation after capture fails before publication",
              persistent.status == "error" and "raw source changed" in persistent.message
              and after_persistent == transient_rows)

        # Restore A, then mutate only after the early post-projection check.  The
        # atomic-save proceed predicate must catch this late boundary and retain A.
        member.write_bytes(b"A")
        real_atomic_save_if = artifact_store.atomic_save_if

        def late_mutation(workbook, path, proceed):
            member.write_bytes(b"B")
            return real_atomic_save_if(workbook, path, proceed)

        with patch(artifact_store, "atomic_save_if", late_mutation):
            late = tsn_library.build_normalized(
                raw, out,
                project=lambda path: (
                    [[Path(path).read_bytes().decode("ascii")]], result),
                **kwargs)
        _title, after_late = workbook_rows(out)
        check("commit-time mutation fails inside atomic publication guard",
              late.status == "error" and "raw source changed" in late.message
              and after_late == transient_rows)

        # The narrowest race occurs after proceed() returns but inside the
        # filesystem replace. The prior bytes may already be replaced here, so
        # the contract is truthful error + no certification, not a false claim
        # that last-good was unconditionally preserved.
        member.write_bytes(b"A")
        real_replace = artifact_store.os.replace
        replaced = []

        def mutate_inside_replace(src, dst):
            if Path(dst) == out and not replaced:
                replaced.append(True)
                member.write_bytes(b"B")
            return real_replace(src, dst)

        with patch(artifact_store.os, "replace", mutate_inside_replace):
            post_predicate = tsn_library.build_normalized(
                raw, out,
                project=lambda path: (
                    [[Path(path).read_bytes().decode("ascii")]], result),
                **kwargs)
        _title, after_post_predicate = workbook_rows(out)
        check("post-predicate mutation cannot return a certified success",
              post_predicate.status == "error" and replaced == [True]
              and "must not be reused" in post_predicate.message
              and after_post_predicate == [["Value"], ["A"]]
              and member.read_bytes() == b"B")
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_reuse_cannot_bypass_admission():
    print("library reuse cannot bypass statewide admission:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_reuse_"))
    try:
        report = "ramp_detail"
        spec = tsn_library.get(report)
        raw = root / report / "raw"
        cons = root / report / "consolidated" / spec.consolidated_name
        raw.mkdir(parents=True)
        cons.parent.mkdir(parents=True)
        first = raw / "one.xlsx"
        first.write_bytes(b"one")
        cons.write_bytes(b"newer-normalized")
        os.utime(first, (1000.0, 1000.0))
        os.utime(cons, (9000.0, 9000.0))
        with patch(tsn_library.paths, "TSN_LIBRARY_ROOT", root):
            legacy = tsn_library.status(report)
            manifest = tsn_library._raw_manifest(report, [first])
            write_tsn_certificate(
                report, cons, raw_manifest={"malformed": True})
            malformed = tsn_library.status(report)
            write_tsn_certificate(report, cons, raw_manifest=manifest)
            accepted = tsn_library.status(report)
            reused = tsn_library.build_consolidated(report)
            first.write_bytes(b"ONE")
            os.utime(first, (1000.0, 1000.0))
            changed_bytes = tsn_library.status(report)
            second = raw / "two.xlsx"
            second.write_bytes(b"two")
            os.utime(second, (2000.0, 2000.0))
            ambiguous = tsn_library.status(report)
        check("one statewide source may reuse a current normalized artifact",
              accepted["raw_admissible"] and accepted["current"]
              and reused.status == "ok" and reused.completion == outcome.COMPLETE
              and reused.skipped_inputs == 0 and reused.failed_inputs == 0)
        check("missing/malformed legacy raw manifest fails reuse closed",
              not legacy["current"] and not legacy["raw_manifest_current"]
              and not malformed["current"] and not malformed["raw_manifest_current"])
        check("a second source makes even a newer normalized artifact non-current",
              ambiguous["raw_count"] == 2 and not ambiguous["raw_admissible"]
              and not ambiguous["current"])
        check("same-count preserved-mtime byte change cannot reuse statewide artifact",
              changed_bytes["raw_count"] == 1 and changed_bytes["raw_admissible"]
              and not changed_bytes["raw_manifest_current"]
              and not changed_bytes["current"])
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_all_registered_reuse_is_content_bound():
    print("all seven registered TSN datasets require exact content manifest for reuse:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_all_manifest_"))
    try:
        with patch(tsn_library.paths, "TSN_LIBRARY_ROOT", root):
            for spec in tsn_library.reports():
                raw = tsn_library.raw_dir(spec.subdir)
                raw.mkdir(parents=True)
                suffix = spec.raw_glob.lstrip("*")
                count = 12 if spec.raw_kind == "district_pdfs" else 1
                members = []
                for number in range(1, count + 1):
                    name = (f"D{number:02d}{suffix}" if count == 12
                            else f"source{suffix}")
                    member = raw / name
                    member.write_bytes(b"A")
                    os.utime(member, (1000.0, 1000.0))
                    members.append(member)
                cons = tsn_library.consolidated_path(spec.subdir)
                cons.parent.mkdir(parents=True)
                cons.write_bytes(b"normalized")
                os.utime(cons, (9000.0, 9000.0))
                manifest = tsn_library._raw_manifest(spec.subdir, members)
                write_tsn_certificate(spec.subdir, cons, raw_manifest=manifest)
                before = tsn_library.status(spec.subdir)
                members[0].write_bytes(b"B")
                os.utime(members[0], (1000.0, 1000.0))
                after = tsn_library.status(spec.subdir)
                check(f"{spec.subdir}: exact bytes reuse, preserved-mtime change stale",
                      before["current"] and before["raw_manifest_current"]
                      and not after["current"] and not after["raw_manifest_current"])
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_normalized_workbook_content_binding():
    print("normalized workbook bytes and canonical TSN identity are durable:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_normalized_identity_"))
    try:
        report = "ramp_detail"
        with patch(tsn_library.paths, "TSN_LIBRARY_ROOT", root):
            raw = tsn_library.raw_dir(report)
            raw.mkdir(parents=True)
            member = raw / "source.xlsx"
            member.write_bytes(b"raw-A")
            os.utime(member, (1000.0, 1000.0))
            out = tsn_library.consolidated_path(report)
            out.parent.mkdir(parents=True)
            out.write_bytes(b"normalized-generation-A")
            os.utime(out, (9000.0, 9000.0))
            manifest = tsn_library._raw_manifest(report)
            identity_a = tsn_library.normalized_workbook_identity(out)
            token_a = tsn_library.canonical_normalized_identity_token(
                report, manifest, identity_a)
            check("canonical identity token is deterministic and domain separated",
                  token_a == tsn_library.canonical_normalized_identity_token(
                      report, manifest, identity_a)
                  and token_a.startswith("tsn-normalized-v1:")
                  and len(token_a.rsplit(":", 1)[-1]) == 64)
            write_tsn_certificate(report, out, raw_manifest=manifest)
            accepted = tsn_library.status(report)
            resolved = tsn_library.resolve(report)
            check("matching raw, catalog version, and normalized bytes are current",
                  accepted["current"] and accepted["normalized_workbook_current"]
                  and accepted["identity_token_current"]
                  and accepted["identity_token"] == token_a
                  and resolved.get("identity_token") == token_a)

            # Replace only normalized bytes while restoring the prior mtime. The
            # legacy mtime/version/raw-only certificate used to accept this.
            previous_mtime = out.stat().st_mtime
            out.write_bytes(b"FORGED-normalized-content")
            os.utime(out, (previous_mtime, previous_mtime))
            replaced = tsn_library.status(report)
            check("preserved-mtime normalized replacement is never reusable",
                  replaced["metadata_current"] and replaced["raw_manifest_current"]
                  and not replaced["normalized_workbook_current"]
                  and not replaced["identity_token_current"]
                  and not replaced["current"])

            rebuilds = []

            def rebuilding_builder(_raw, target, events=None,
                                   confirm_overwrite=None):
                del events, confirm_overwrite
                rebuilds.append(True)
                Path(target).write_bytes(b"normalized-generation-rebuilt")
                built = ConsolidateResult(
                    status="ok", output_path=str(target),
                    completion=outcome.COMPLETE, skipped_inputs=0, failed_inputs=0)
                built.tsn_raw_manifest = tsn_library._raw_manifest(report)
                return built

            class RebuildModule:
                build_into = staticmethod(rebuilding_builder)

            with patch(tsn_library.importlib, "import_module",
                       lambda _name: RebuildModule):
                healed = tsn_library.ensure_current(report)
            rebuilt = tsn_library.status(report)
            check("ensure_current rebuilds instead of reading replaced bytes",
                  rebuilds == [True] and healed.status == "ok" and rebuilt["current"]
                  and rebuilt["identity_token"] != token_a)

            # Legacy/malformed workbook fields fail closed even when every older
            # provenance field and mtime still agrees.
            identity_forged = tsn_library.normalized_workbook_identity(out)
            tsn_library.consolidation_meta.write_outcome(
                out, ConsolidateResult(status="ok", completion=outcome.COMPLETE),
                extra={"tsn_normalization_version": tsn_library.get(
                           report).normalization_version,
                       "tsn_raw_manifest": manifest})
            legacy = tsn_library.status(report)
            write_tsn_certificate(
                report, out, raw_manifest=manifest,
                workbook_identity={"version": 1, "algorithm": "sha256",
                                   "byte_length": True, "sha256": "0" * 64})
            malformed = tsn_library.status(report)
            write_tsn_certificate(
                report, out, raw_manifest=manifest,
                workbook_identity=identity_forged,
                identity_token="tsn-normalized-v1:" + "0" * 64)
            wrong_token = tsn_library.status(report)
            tsn_library.consolidation_meta.meta_path(out).write_text(
                '{"schema_version":1,"schema_version":1}', encoding="utf-8")
            duplicate_key = tsn_library.status(report)
            check("missing, malformed, or mismatched workbook provenance fails closed",
                  not legacy["current"] and not legacy["normalized_workbook_current"]
                  and not malformed["current"] and not malformed["normalized_workbook_current"]
                  and not wrong_token["current"] and not wrong_token["identity_token_current"]
                  and not duplicate_key["current"]
                  and "strict UTF-8 JSON" in duplicate_key["certificate_error"])
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_build_rereads_exact_normalized_bytes():
    print("build success requires durable sidecar reread against exact workbook bytes:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_build_reread_"))
    try:
        report = "ramp_detail"
        with patch(tsn_library.paths, "TSN_LIBRARY_ROOT", root):
            raw = tsn_library.raw_dir(report)
            raw.mkdir(parents=True)
            (raw / "source.xlsx").write_bytes(b"raw-A")
            out = tsn_library.consolidated_path(report)

            def certified_builder(_raw, target, events=None, confirm_overwrite=None):
                del events, confirm_overwrite
                Path(target).parent.mkdir(parents=True, exist_ok=True)
                Path(target).write_bytes(b"normalized-generation-A")
                built = ConsolidateResult(
                    status="ok", output_path=str(target),
                    completion=outcome.COMPLETE, skipped_inputs=0, failed_inputs=0)
                built.tsn_raw_manifest = tsn_library._raw_manifest(report)
                return built

            class FakeModule:
                build_into = staticmethod(certified_builder)

            real_write = tsn_library.consolidation_meta.write_outcome
            tampered = []

            def publish_then_replace(workbook, result, extra=None, commit_guard=None):
                published = real_write(
                    workbook, result, extra=extra, commit_guard=commit_guard)
                if (published and result.status == "ok"
                        and result.completion == outcome.COMPLETE and not tampered):
                    timestamp = Path(workbook).stat().st_mtime
                    Path(workbook).write_bytes(b"normalized-generation-B")
                    os.utime(workbook, (timestamp, timestamp))
                    tampered.append(True)
                return published

            with patch(tsn_library.importlib, "import_module", lambda _name: FakeModule), \
                 patch(tsn_library.consolidation_meta, "write_outcome", publish_then_replace):
                result = tsn_library.build_consolidated(report, force=True)
            state = tsn_library.status(report)
            check("post-publication byte replacement makes the build return error",
                  tampered == [True] and result.status == "error"
                  and "exact normalized bytes" in result.message
                  and not state["normalized_workbook_current"] and not state["current"])
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_build_time_source_change_is_uncertified():
    print("source mutation during a statewide build is durably uncertified:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_build_mutation_"))
    try:
        report = "ramp_detail"
        with patch(tsn_library.paths, "TSN_LIBRARY_ROOT", root):
            raw = tsn_library.raw_dir(report)
            raw.mkdir(parents=True)
            member = raw / "source.xlsx"
            member.write_bytes(b"A")
            os.utime(member, (1000.0, 1000.0))

            def mutating_builder(_raw, out, events=None, confirm_overwrite=None):
                del events, confirm_overwrite
                Path(out).write_bytes(b"normalized-from-A")
                member.write_bytes(b"B")
                os.utime(member, (1000.0, 1000.0))
                return ConsolidateResult(
                    status="ok", output_path=str(out), completion=outcome.COMPLETE)

            class FakeModule:
                build_into = staticmethod(mutating_builder)

            with patch(tsn_library.importlib, "import_module", lambda _name: FakeModule):
                result = tsn_library.build_consolidated(report, force=True)
            state = tsn_library.status(report)
            check("changed pre/post manifest returns error",
                  result.status == "error" and "changed during normalization" in result.message)
            check("already-written unstable workbook cannot become reusable",
                  not state["current"] and not state["raw_manifest_current"])
            check("unstable generation is durably marked partial when retained",
                  (not tsn_library.consolidated_path(report).exists()
                   or tsn_library.consolidation_meta.read_completion(
                       tsn_library.consolidated_path(report)) == outcome.PARTIAL))
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_complete_sidecar_failure_is_not_tsn_success():
    print("complete TSN builds require a durably readable source certificate:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_sidecar_failure_"))
    try:
        report = "ramp_detail"
        with patch(tsn_library.paths, "TSN_LIBRARY_ROOT", root):
            raw = tsn_library.raw_dir(report)
            raw.mkdir(parents=True)
            member = raw / "source.xlsx"
            member.write_bytes(b"A")
            out = tsn_library.consolidated_path(report)

            def certified_builder(_raw, target, events=None, confirm_overwrite=None):
                del events, confirm_overwrite
                wb = Workbook()
                wb.active.title = "Normalized"
                wb.active.append(["Value"])
                wb.active.append(["A"])
                Path(target).parent.mkdir(parents=True, exist_ok=True)
                wb.save(target)
                wb.close()
                built = ConsolidateResult(
                    status="ok", output_path=str(target),
                    completion=outcome.COMPLETE, skipped_inputs=0, failed_inputs=0)
                built.tsn_raw_manifest = tsn_library._raw_manifest(report)
                return built

            class FakeModule:
                build_into = staticmethod(certified_builder)

            real_replace = tsn_library.consolidation_meta.os.replace
            sidecar = tsn_library.consolidation_meta.meta_path(out)

            def deny_sidecar(src, dst):
                if Path(dst) == sidecar:
                    raise PermissionError("forced TSN certificate publication failure")
                return real_replace(src, dst)

            with patch(tsn_library.importlib, "import_module", lambda _name: FakeModule), \
                 patch(tsn_library.consolidation_meta.os, "replace", deny_sidecar):
                result = tsn_library.build_consolidated(report, force=True)
            state = tsn_library.status(report)
            check("generic COMPLETE-sidecar fallback cannot certify a TSN build",
                  result.status == "error" and "durable TSN" in result.message
                  and "certificate" in result.message)
            check("unpublished certificate remains non-current and non-reusable",
                  out.exists() and not sidecar.exists()
                  and not state["metadata_current"]
                  and not state["raw_manifest_current"] and not state["current"])
    finally:
        shutil.rmtree(root, ignore_errors=True)


def _mutations(header):
    missing = list(header)
    missing.pop(len(missing) // 2)
    duplicate = list(header) + [header[0]]
    reordered = list(header)
    reordered[0], reordered[1] = reordered[1], reordered[0]
    renamed = list(header)
    renamed[-1] = f"{renamed[-1]}_RENAMED"
    return (("missing", missing), ("duplicate", duplicate),
            ("reordered", reordered), ("renamed", renamed))


def test_detail_headers():
    print("detail raw schemas reject drift before projection (normalizer + direct loader):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_headers_"))
    try:
        cases = (
            ("Ramp Detail", rd, rd_load, rd.TSN_RAW_HEADER, "_tsn_raw_row"),
            ("Intersection Detail", idt, id_load, idt.TSN_RAW_HEADER, "_tsn_row"),
            ("Highway Detail", hdt, hd_load, hdt.TSN_RAW_HEADER, "_tsn_row"),
        )
        for label, comparator, loader, header, projector_name in cases:
            for mutation, changed in _mutations(header):
                path = root / f"{label}_{mutation}.xlsx"
                write_raw(path, changed)
                calls = []

                def forbidden(*_args, **_kwargs):
                    calls.append("projected")
                    return []

                errors = []
                with patch(comparator, projector_name, forbidden):
                    for fn in (loader.tsn_rows_with_dcr, comparator.tsn_rows_from_raw):
                        try:
                            fn(path)
                        except ValueError as exc:
                            errors.append(str(exc))
                check(f"{label}: {mutation} header rejected on both raw paths",
                      len(errors) == 2 and all("raw header does not match" in e for e in errors))
                check(f"{label}: {mutation} rejected before a row projector runs", not calls)
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_detail_document_and_row_claims():
    print("detail raw documents reject alternate/extra sheets, formulas/errors, and blank identities:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_documents_"))
    try:
        cases = (
            ("Ramp Detail", rd, rd_load, rd.TSN_RAW_HEADER,
             {"LOCATION": "01-DN-001", "PR": "R", "PM": 1.0}),
            ("Intersection Detail", idt, id_load, idt.TSN_RAW_HEADER,
             {"LOCATION": "12 ORA 001", "POST_MILE": 1.0}),
            ("Highway Detail", hdt, hd_load, hdt.TSN_RAW_HEADER,
             {"DIST": "01", "CNTY": "DN", "RTE": 1, "POSTMILE": 1.0}),
        )
        for label, comparator, loader, header, identity in cases:
            def source_row():
                row = [None] * len(header)
                for name, value in identity.items():
                    row[list(header).index(name)] = value
                return row

            variants = []
            renamed = root / f"{label}_renamed.xlsx"
            write_raw(renamed, header, source_row())
            wb = load_workbook(renamed)
            wb.active.title = "Renamed"
            wb.save(renamed)
            wb.close()
            variants.append(("renamed sole sheet", renamed, "exactly one worksheet"))

            for state in ("visible", "hidden"):
                extra = root / f"{label}_extra_{state}.xlsx"
                write_raw(extra, header, source_row())
                wb = load_workbook(extra)
                ws = wb.create_sheet("Extra")
                ws.sheet_state = state
                wb.save(extra)
                wb.close()
                variants.append((f"extra {state} sheet", extra, "exactly one worksheet"))

            for kind, value, data_type in (("formula", "=1+1", "f"),
                                           ("error", "#DIV/0!", "e")):
                bad = root / f"{label}_{kind}.xlsx"
                write_raw(bad, header, source_row())
                wb = load_workbook(bad)
                cell = wb.active.cell(row=2, column=len(header))
                cell.value = value
                cell.data_type = data_type
                wb.save(bad)
                wb.close()
                variants.append((f"{kind} data cell", bad, "formula/error cell"))

            missing = root / f"{label}_missing_identity.xlsx"
            row = source_row()
            missing_name = next(iter(identity))
            row[list(header).index(missing_name)] = None
            write_raw(missing, header, row)
            variants.append(("blank required identity", missing,
                             "missing required identity claim"))

            for variant, path, fragment in variants:
                errors = []
                for fn in (loader.tsn_rows_with_dcr, comparator.tsn_rows_from_raw):
                    try:
                        fn(path)
                    except ValueError as exc:
                        errors.append(str(exc))
                check(f"{label}: {variant} rejected on both raw paths",
                      len(errors) == 2 and all(fragment in error for error in errors))
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_detail_success_outcomes():
    print("detail successful builders own COMPLETE state and emit admitted rows:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_complete_"))
    try:
        cases = (
            ("Ramp Detail", rd, rd_load, rd.TSN_RAW_HEADER, "_tsn_raw_row",
             "Ramp Detail (TSN)", ["Route"] + rd.SHARED_HEADER + rd_load.SIDECAR_HEADER,
             ["001"] + [f"rd{i}" for i in range(len(rd.SHARED_HEADER))]),
            ("Intersection Detail", idt, id_load, idt.TSN_RAW_HEADER, "_tsn_row",
             "Intersection Detail (TSN)", ["Route"] + idt.SHARED_HEADER + id_load.SIDECAR_HEADER,
             ["001"] + [f"id{i}" for i in range(len(idt.SHARED_HEADER))]),
            ("Highway Detail", hdt, hd_load, hdt.TSN_RAW_HEADER, "_tsn_row",
             hdt.NORMALIZED_SHEET, ["Route"] + hdt.SHARED_HEADER + hd_load.SIDECAR_HEADER,
             ["001"] + [f"hd{i}" for i in range(len(hdt.SHARED_HEADER))]),
        )
        for label, comparator, loader, header, projector_name, sheet, expected_header, base_row in cases:
            raw = root / label.replace(" ", "_")
            raw.mkdir()
            source = raw / "truth.xlsx"
            source_row = ["x"] + [None] * (len(header) - 1)
            # Feed the sidecar extractors recognizable source claims.
            for name, value in (("LOCATION", "12 ORA 001" if label.startswith("Intersection")
                                  else "01-DN-001"), ("DIST", "01"), ("CNTY", "DN")):
                if name in header:
                    source_row[list(header).index(name)] = value
            required_values = ({"PR": "R", "PM": 1.0} if label.startswith("Ramp")
                               else {"POST_MILE": 1.0} if label.startswith("Intersection")
                               else {"RTE": 1, "POSTMILE": 1.0})
            for name, value in required_values.items():
                source_row[list(header).index(name)] = value
            write_raw(source, header, source_row)
            out = root / f"{label.replace(' ', '_')}.xlsx"
            with patch(comparator, projector_name,
                       lambda _r, _h, row=list(base_row): list(row)):
                result = loader.build_into(raw, out)
            title, rows = workbook_rows(out)
            check(f"{label}: success explicitly returns complete with no rejected inputs",
                  result.status == "ok" and result.completion == outcome.COMPLETE
                  and result.skipped_inputs == 0 and result.failed_inputs == 0)
            check(f"{label}: one ordinary source independently admitted",
                  tsn_library._statewide_raw_candidates(raw, "*.xlsx") == [source])
            check(f"{label}: exact normalized header and exactly one emitted row",
                  title == sheet and rows[0] == expected_header and len(rows[1:]) == 1
                  and rows[1][:-2] == base_row)
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_summary_exact_one():
    print("Ramp/Intersection Summary use the same exact-one statewide contract:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_tsn_summary_admission_"))
    try:
        cases = (
            ("Ramp Summary", rs_load, rstsn, list(rstsn._CATEGORIES), "total_ramps"),
            ("Intersection Summary", is_load, istsn,
             list(istsn._SPEC.categories_for("tsn")), "total_intersections"),
        )
        for label, loader, comparator, categories, total_slug in cases:
            raw = root / label.replace(" ", "_")
            raw.mkdir()
            first = raw / "a.pdf"
            second = raw / "b.pdf"
            first.write_bytes(b"a")
            second.write_bytes(b"b")
            (raw / "~$owner.pdf").write_bytes(b"lock")
            calls = []
            with patch(comparator, "parse_tsn_pdf", lambda p: calls.append(p) or {}):
                ambiguous = loader.build_into(raw, root / f"{label}_ambiguous.xlsx")
            check(f"{label}: two ordinary PDFs rejected before parse",
                  ambiguous.status == "error" and "Found 2 ordinary matching files" in ambiguous.message
                  and not calls)

            second.unlink()
            counts = {slug: i + 1 for i, (_key, slug) in enumerate(categories)}
            counts[total_slug] = 999
            out = root / f"{label}_one.xlsx"
            with patch(comparator, "parse_tsn_pdf", lambda _p, c=counts: dict(c)):
                accepted = loader.build_into(raw, out)
            _title, rows = workbook_rows(out)
            check(f"{label}: one PDF plus owner-lock succeeds COMPLETE",
                  accepted.status == "ok" and accepted.completion == outcome.COMPLETE
                  and accepted.skipped_inputs == 0 and accepted.failed_inputs == 0
                  and len(rows[1:]) == len(categories))
    finally:
        shutil.rmtree(root, ignore_errors=True)


def main():
    test_shared_exact_one()
    test_statewide_immutable_snapshot_and_commit_guard()
    test_reuse_cannot_bypass_admission()
    test_all_registered_reuse_is_content_bound()
    test_normalized_workbook_content_binding()
    test_build_rereads_exact_normalized_bytes()
    test_build_time_source_change_is_uncertified()
    test_complete_sidecar_failure_is_not_tsn_success()
    test_detail_headers()
    test_detail_document_and_row_claims()
    test_detail_success_outcomes()
    test_summary_exact_one()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL TSN RAW-SOURCE CONTRACT CHECKS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
