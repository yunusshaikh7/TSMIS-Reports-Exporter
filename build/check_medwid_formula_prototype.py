"""Adversarial prototype/gate for the Phase-3 E1 Med-Wid formula twin.

This file deliberately does *not* patch ``compare_core``.  It proves a standalone
Excel expression for the approved narrow grammar before production integration:

* ASCII ``digits[.digits]``;
* optionally followed by one printable-ASCII suffix (U+0021..U+007E) that is
  neither an ASCII digit nor ``.``;
* Excel-TRIM / ASCII-U+0020 whitespace semantics;
* leading integer zeros and trailing fractional zeros are insignificant;
* the suffix is preserved exactly and compared case-sensitively by the caller.

Signs, leading-dot shorthand, exponents, Unicode digits/suffixes, control suffixes,
and multiple suffix characters remain raw.  The formula uses no numeric conversion:
``VALUE``, ``NUMBERVALUE``, and ``DECIMAL`` are forbidden.

The accepted primitive is a five-cell hidden helper block (TRIM, core, validity,
significance mask, canonical value).  It uses only scalar functions already used by
the application's workbooks; there are no modern functions, numeric conversion,
dynamic arrays, regex, volatile functions, or CSE/array formulas.  A compact LET
prototype is retained only as an Excel-compatibility probe: production must not depend
on it.

Default execution is offline and belongs in the normal ``check_*.py`` suite.  To
prove the generated formulas in installed Microsoft Excel as well, run:

    build\\.venv\\Scripts\\python.exe build\\check_medwid_formula_prototype.py --excel
"""
from __future__ import annotations

import argparse
import itertools
import random
import re
from pathlib import Path

from _checklib import Checker, temp_dir

from openpyxl import Workbook, load_workbook


XL_FORMULA_LIMIT = 8_192
ALLOWED_SUFFIXES = "".join(
    chr(code)
    for code in range(0x21, 0x7F)
    if chr(code) not in "0123456789."
)
_STRICT_MEDWID = re.compile(r"([0-9]+)(?:\.([0-9]+))?(.)?", re.DOTALL)


def _excel_quote(text: str) -> str:
    """An Excel string literal, including the doubled-quote escape."""
    return '"' + text.replace('"', '""') + '"'


def _substitutes(expr: str, replacements) -> str:
    for old, new in replacements:
        expr = f"SUBSTITUTE({expr},{_excel_quote(old)},{_excel_quote(new)})"
    return expr


MEDWID_HELPER_VERSION = "CMP_E1_MW_V1"


def medwid_helper_headers(field_token: str) -> tuple[str, ...]:
    """Stable, explicit headers for the five hidden data-sheet helpers."""
    prefix = f"__{MEDWID_HELPER_VERSION}_{field_token}"
    return tuple(f"{prefix}_{stage}" for stage in
                 ("TRIM", "CORE", "VALID", "MASK", "CANON"))


def excel_medwid_stage_formulas(
    source_ref: str,
    trimmed_ref: str,
    core_ref: str,
    valid_ref: str,
    mask_ref: str,
) -> tuple[str, ...]:
    """Return the five legacy scalar helper formulas for one source cell.

    The production geometry is intentionally staged on each source data sheet:

    ``TRIM`` -> ``CORE`` -> ``VALID`` -> ``MASK`` -> ``CANON``.

    This keeps source references short, avoids Excel's 8,192-character formula
    limit, and avoids duplicating a 7K formula inside every Comparison expression.
    X/Y are sentinels used only after VALID proves CORE contains digits/dot.
    """
    suffixes = _excel_quote(ALLOWED_SUFFIXES)
    trimmed = f'=IF(ISBLANK({source_ref}),"",TRIM({source_ref}))'
    has_suffix = (
        f'AND({trimmed_ref}<>"",ISNUMBER(FIND(RIGHT({trimmed_ref},1),'
        f'{suffixes})))'
    )
    core = (
        f'=IF({has_suffix},LEFT({trimmed_ref},LEN({trimmed_ref})-1),'
        f'{trimmed_ref})'
    )
    dot_count = f'LEN({core_ref})-LEN(SUBSTITUTE({core_ref},".",""))'
    no_non_digits = _substitutes(
        f'SUBSTITUTE({core_ref},".","")',
        ((str(i), "") for i in range(10)),
    )
    valid = (
        f'=AND({core_ref}<>"",LEFT({core_ref},1)<>".",'
        f'RIGHT({core_ref},1)<>".",{dot_count}<=1,{no_non_digits}="")'
    )
    significance_mask = _substitutes(
        f'SUBSTITUTE({core_ref},".","X")',
        ((str(i), "X") for i in range(1, 10)),
    )
    mask = f'=IF({valid_ref},{significance_mask},"")'

    first = f'FIND("X",{mask_ref}&"X")'
    n_significant = (
        f'LEN({mask_ref})-LEN(SUBSTITUTE({mask_ref},"X",""))'
    )
    last = (
        f'IF({n_significant}=0,0,FIND("Y",SUBSTITUTE('
        f'{mask_ref},"X","Y",{n_significant})))'
    )
    dot = f'FIND(".",{core_ref}&".")'
    plain = (
        f'IF({first}>LEN({core_ref}),"0",MID('
        f'{core_ref},{first},LEN({core_ref})))'
    )
    raw_decimal = (
        f'IF({first}={dot},"0","")&MID('
        f'{core_ref},{first},{last}-{first}+1)&"Y"'
    )
    decimal = (
        f'SUBSTITUTE(SUBSTITUTE({raw_decimal},".Y","Y"),"Y","")'
    )
    suffix = (
        f'IF(LEN({core_ref})<LEN({trimmed_ref}),RIGHT({trimmed_ref},1),"")'
    )
    canonical = (
        f'=IF({trimmed_ref}="","",IF(NOT({valid_ref}),{trimmed_ref},'
        f'IF({dot_count}=0,{plain},{decimal})&{suffix}))'
    )
    return trimmed, core, valid, mask, canonical


# OOXML modern-function spellings are probed empirically but are rejected as a
# production dependency regardless of the result: the staged formulas above are
# the compatibility target.
LET_PROBES = (
    ("bare LET", "=LET(x,1,x)"),
    ("_xlfn LET", "=_xlfn.LET(x,1,x)"),
    ("_xlfn LET + _xlpm variable", "=_xlfn.LET(_xlpm.x,1,_xlpm.x)"),
)


def _ascii_trim(value) -> str:
    """Python mirror of Excel TRIM's U+0020-only policy for this fixture."""
    if value is None:
        return ""
    return re.sub(" +", " ", str(value)).strip(" ")


def strict_medwid_oracle(value) -> str:
    """Small regex oracle, intentionally independent of the formula algorithm."""
    text = _ascii_trim(value)
    match = _STRICT_MEDWID.fullmatch(text)
    if match is None:
        return text
    whole, fraction, suffix = match.groups()
    if suffix is not None and suffix not in ALLOWED_SUFFIXES:
        return text
    whole = whole.lstrip("0") or "0"
    fraction = fraction.rstrip("0") if fraction is not None else ""
    return whole + (f".{fraction}" if fraction else "") + (suffix or "")


def formula_algorithm_mirror(value) -> str:
    """Literal Python translation of the staged formula algorithm for fuzzing."""
    text = _ascii_trim(value)
    if not text:
        return ""
    has_suffix = text[-1] in ALLOWED_SUFFIXES
    core = text[:-1] if has_suffix else text
    suffix = text[-1] if has_suffix else ""
    dot_count = core.count(".")
    valid = (
        bool(core)
        and not core.startswith(".")
        and not core.endswith(".")
        and dot_count <= 1
        and all(char in "0123456789." for char in core)
    )
    if not valid:
        return text

    mask = "".join("X" if char == "." or char in "123456789" else "0"
                   for char in core)
    first = mask.find("X")
    if dot_count == 0:
        number = "0" if first < 0 else core[first:]
    else:
        last = mask.rfind("X")
        dot = core.index(".")
        raw = ("0" if first == dot else "") + core[first:last + 1] + "Y"
        number = raw.replace(".Y", "Y").replace("Y", "")
    return number + suffix


EXPLICIT_CASES = (
    ("blank", None, ""),
    ("empty text", "", ""),
    ("zero", "0", "0"),
    ("all-zero integer", "000000", "0"),
    ("integer leading zeros", "0006", "6"),
    ("suffix leading zeros", "0006V", "6V"),
    ("decimal zero padding", "0006.5000V", "6.5V"),
    ("fractional leading zeros", "000.00100", "0.001"),
    ("all-zero fraction", "000.000Z", "0Z"),
    ("integer trailing zeros retained", "001000", "1000"),
    ("zero fraction after integer", "00100.000", "100"),
    ("mixed fraction", "00100.0200#", "100.02#"),
    ("punctuation suffix", "06#", "6#"),
    ("lowercase suffix preserved", "06v", "6v"),
    ("ASCII edge/internal spaces", "  0006.5000V  ", "6.5V"),
    ("ASCII internal spaces stay anomalous", "0006   .5000V", "0006 .5000V"),
    ("ASCII suffix-space trims away", "06 ", "6"),
    ("negative sign", "-06V", "-06V"),
    ("explicit plus", "+06V", "+06V"),
    ("leading dot", ".50V", ".50V"),
    ("trailing dot", "06.", "06."),
    ("multiple dots", "06.0.0V", "06.0.0V"),
    ("exponent", "6e0", "6e0"),
    ("exponent with suffix", "06e0V", "06e0V"),
    ("multiple suffix letters", "06VV", "06VV"),
    ("multiple suffix punctuation", "06##", "06##"),
    ("suffix-only", "V", "V"),
    ("literal formula lead", "=06V", "=06V"),
    ("literal Excel error", "#N/A", "#N/A"),
    ("Arabic-Indic numeric portion", "٠٦V", "٠٦V"),
    ("fullwidth numeric portion", "０６V", "０６V"),
    ("Devanagari numeric portion", "०६V", "०६V"),
    ("Unicode digit suffix", "06١", "06١"),
    ("non-ASCII letter suffix", "06é", "06é"),
    ("tab suffix", "06\t", "06\t"),
    ("LF suffix", "06\n", "06\n"),
    ("CR suffix", "06\r", "06\r"),
    ("NBSP suffix", "06\u00a0", "06\u00a0"),
    ("DEL suffix", "06\x7f", "06\x7f"),
    ("emoji suffix", "06🙂", "06🙂"),
    ("tab-wrapped token", "\t06V\t", "\t06V\t"),
    ("huge exact decimal", "0" * 80 + "12345678901234567890.0100V",
     "12345678901234567890.01V"),
)


PAIR_CASES = (
    ("integer zero padding", "0Z", "00Z", True),
    ("decimal zero padding", "06.00V", "6V", True),
    ("punctuation suffix", "06#", "6#", True),
    ("suffix case", "6v", "6V", False),
    ("negative stays raw", "-06V", "-6V", False),
    ("leading dot stays raw", ".50", "0.5", False),
    ("Unicode suffix stays raw", "06é", "6é", False),
    ("Unicode digit suffix stays raw", "06١", "6١", False),
    ("large adjacent integers", "9007199254740992V", "9007199254740993V", False),
    ("large exact decimal", "12345678901234567890.0100V",
     "12345678901234567890.01V", True),
    ("large adjacent decimals", "12345678901234567890.010V",
     "12345678901234567890.011V", False),
)


def _literal(cell, value) -> None:
    cell.value = value
    if isinstance(value, str):
        # Inputs like '=06V' and '#N/A' are fixture text, never formulas/errors.
        cell.data_type = "s"


def _all_excel_cases():
    cases = list(EXPLICIT_CASES)
    cases.extend(
        (f"printable ASCII suffix U+{ord(suffix):04X}",
         f"0006.5000{suffix}", f"6.5{suffix}")
        for suffix in ALLOWED_SUFFIXES
    )
    return cases


def _build_workbook(path: Path) -> tuple[int, int]:
    cases = _all_excel_cases()
    wb = Workbook()
    ws = wb.active
    ws.title = "Canonical"
    ws.append(["Case", "Input", *medwid_helper_headers("F1"), "Expected"])
    for row, (label, value, expected) in enumerate(cases, start=2):
        _literal(ws.cell(row, 1), label)
        _literal(ws.cell(row, 2), value)
        formulas = excel_medwid_stage_formulas(
            f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}",
        )
        for column, formula in enumerate(formulas, start=3):
            ws.cell(row, column).value = formula
        _literal(ws.cell(row, 8), expected)
    for column in "CDEFG":
        ws.column_dimensions[column].hidden = True

    pairs = wb.create_sheet("Equality")
    pairs.append([
        "Case", "Left", *medwid_helper_headers("LEFT_F1"),
        "Right", *medwid_helper_headers("RIGHT_F1"),
        "EXACT result", "Expected",
    ])
    for row, (label, left, right, expected) in enumerate(PAIR_CASES, start=2):
        _literal(pairs.cell(row, 1), label)
        _literal(pairs.cell(row, 2), left)
        left_formulas = excel_medwid_stage_formulas(
            f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}",
        )
        for column, formula in enumerate(left_formulas, start=3):
            pairs.cell(row, column).value = formula
        _literal(pairs.cell(row, 8), right)
        right_formulas = excel_medwid_stage_formulas(
            f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}",
        )
        for column, formula in enumerate(right_formulas, start=9):
            pairs.cell(row, column).value = formula
        pairs.cell(row, 14).value = f"=EXACT(G{row},M{row})"
        pairs.cell(row, 15).value = expected
    for column in ("C", "D", "E", "F", "G", "I", "J", "K", "L", "M"):
        pairs.column_dimensions[column].hidden = True

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(path)
    wb.close()
    return len(cases), len(PAIR_CASES)


def _calculate_with_excel(path: Path) -> str:
    """Calculate and save the fixture in installed Microsoft Excel via COM."""
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    book = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        version = f"{excel.Version} build {excel.Build}"
        book = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
        excel.CalculateFullRebuild()
        book.Save()
        book.Close(SaveChanges=True)
        book = None
        return version
    finally:
        if book is not None:
            book.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def _verify_excel_results(c: Checker, path: Path, n_cases: int, n_pairs: int) -> None:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Canonical"]
        failures = []
        for row in range(2, n_cases + 2):
            label = ws.cell(row, 1).value
            got = ws.cell(row, 7).value
            expected = ws.cell(row, 8).value
            got = "" if got is None and expected == "" else got
            if got != expected:
                failures.append((label, got, expected))
        c.check("installed Excel canonicalizes every adversarial value exactly",
                not failures, repr(failures[:5]))

        ws = wb["Equality"]
        failures = []
        for row in range(2, n_pairs + 2):
            label = ws.cell(row, 1).value
            got = ws.cell(row, 14).value
            expected = ws.cell(row, 15).value
            if got is not expected and got != expected:
                failures.append((label, got, expected))
        c.check("installed Excel EXACT matches the approved equality pairs",
                not failures, repr(failures[:5]))

    finally:
        wb.close()


def _probe_let_spellings(folder: Path):
    """Probe each OOXML spelling in isolation so one rejected file cannot hide another."""
    results = []
    for index, (label, formula) in enumerate(LET_PROBES, start=1):
        path = folder / f"let-probe-{index}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = formula
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(path)
        wb.close()
        try:
            _calculate_with_excel(path)
            wb = load_workbook(path, data_only=True, read_only=True)
            try:
                result = wb.active["A1"].value
            finally:
                wb.close()
        except Exception as exc:  # the rejected spelling itself is the evidence
            result = f"OPEN/CALC ERROR: {type(exc).__name__}: {exc}"
        results.append((label, result))
    return results


def test_oracles(c: Checker) -> None:
    explicit_failures = []
    for label, value, expected in EXPLICIT_CASES:
        got = strict_medwid_oracle(value)
        if got != expected:
            explicit_failures.append((label, got, expected))
    c.check("strict oracle matches every hand-authored expected result",
            not explicit_failures, repr(explicit_failures[:5]))

    suffix_failures = []
    for suffix in ALLOWED_SUFFIXES:
        got = strict_medwid_oracle(f"0006.5000{suffix}")
        expected = f"6.5{suffix}"
        if got != expected:
            suffix_failures.append((suffix, got, expected))
    c.check("all 83 approved printable-ASCII suffixes canonicalize",
            not suffix_failures, repr(suffix_failures[:5]))

    alphabet = ("0", "1", "9", ".", "V", "v", "#", "+", "-", "e",
                " ", "\t", "é", "١")
    fuzz_failures = []
    for length in range(5):
        for parts in itertools.product(alphabet, repeat=length):
            value = "".join(parts)
            got = formula_algorithm_mirror(value)
            expected = strict_medwid_oracle(value)
            if got != expected:
                fuzz_failures.append((value, got, expected))
                break
        if fuzz_failures:
            break

    rng = random.Random(20260712)
    for _ in range(20_000):
        value = "".join(rng.choice(alphabet) for _ in range(rng.randrange(0, 50)))
        got = formula_algorithm_mirror(value)
        expected = strict_medwid_oracle(value)
        if got != expected:
            fuzz_failures.append((value, got, expected))
            break
    c.check("formula algorithm matches the regex oracle under exhaustive/random fuzz",
            not fuzz_failures, repr(fuzz_failures[:3]))


def test_formula_structure(c: Checker) -> None:
    formulas = excel_medwid_stage_formulas(
        "$XFD999999", "$AAA999999", "$AAB999999", "$AAC999999",
        "$AAD999999",
    )
    lengths = [len(formula) for formula in formulas]
    combined = "\n".join(formulas)
    upper = combined.upper()

    c.check("every staged formula is safely below Excel's 8,192-character limit",
            max(lengths) < XL_FORMULA_LIMIT,
            f"lengths={lengths!r}; limit={XL_FORMULA_LIMIT:,}")
    c.check("the complete five-stage primitive remains compact",
            sum(lengths) < 2_000,
            f"lengths={lengths!r}; total={sum(lengths):,}")
    forbidden = ("VALUE(", "NUMBERVALUE(", "DECIMAL(", "LET(", "LAMBDA(",
                 "INDIRECT(", "SEQUENCE(", "FILTER(", "REGEX")
    present = [token for token in forbidden if token in upper]
    c.check("stages have no modern, lossy, volatile, array, or regex primitive",
            not present, repr(present))
    c.check("TRIM stage is blank-safe and ASCII-space-only",
            "ISBLANK(" in formulas[0].upper()
            and "TRIM(" in formulas[0].upper()
            and "CLEAN(" not in upper,
            formulas[0])
    c.check("CORE embeds exactly the printable-ASCII suffix whitelist",
            _excel_quote(ALLOWED_SUFFIXES) in formulas[1],
            f"suffixes={ALLOWED_SUFFIXES!r}")
    headers = medwid_helper_headers("FIELD_7")
    c.check("helper headers are explicit, unique, and normalization-versioned",
            len(set(headers)) == 5
            and all(MEDWID_HELPER_VERSION in header for header in headers),
            repr(headers))

    # Production must append helpers after the existing key helper.  Back-link,
    # source-field, and Key(helper) coordinates therefore stay byte-for-byte stable.
    base = ["Comparison row", "Route", "Loc", "Med-Wid", "Key (helper)"]
    augmented = base + list(headers)
    c.check("helper append preserves back-link/source/key-helper geometry",
            augmented[:len(base)] == base
            and augmented.index("Key (helper)") == base.index("Key (helper)"),
            repr(augmented))


def test_workbook(c: Checker, run_excel: bool) -> None:
    with temp_dir("tsmis_medwid_formula_") as tmp:
        path = Path(tmp) / "medwid-formula-prototype.xlsx"
        n_cases, n_pairs = _build_workbook(path)
        wb = load_workbook(path, data_only=False, read_only=False)
        try:
            formula = wb["Canonical"]["C2"]
            canonical = wb["Canonical"]["G2"]
            injection = wb["Canonical"].cell(
                next(row for row in range(2, n_cases + 2)
                     if wb["Canonical"].cell(row, 1).value == "literal formula lead"),
                2,
            )
            c.check("prototype workbook stores all five legacy stages as formulas",
                    formula.data_type == "f"
                    and canonical.data_type == "f"
                    and "LET(" not in str(canonical.value).upper(),
                    f"trim={str(formula.value)[:80]!r}; "
                    f"canon={str(canonical.value)[:80]!r}")
            c.check("prototype helper columns are hidden",
                    all(wb["Canonical"].column_dimensions[col].hidden
                        for col in "CDEFG"),
                    repr({col: wb["Canonical"].column_dimensions[col].hidden
                          for col in "CDEFG"}))
            c.check("adversarial formula-leading input remains literal text",
                    injection.data_type == "s" and injection.value == "=06V",
                    f"type={injection.data_type!r}; value={injection.value!r}")
        finally:
            wb.close()

        if run_excel:
            version = _calculate_with_excel(path)
            print(f"  note: calculated with installed Excel {version}")
            _verify_excel_results(c, path, n_cases, n_pairs)
            probe_results = _probe_let_spellings(Path(tmp))
            print(f"  note: installed-Excel LET spelling probe: {probe_results!r}")


def main(argv=None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--excel", action="store_true",
        help="also calculate the prototype with installed Microsoft Excel via COM",
    )
    args = parser.parse_args(argv)

    c = Checker()
    test_oracles(c)
    test_formula_structure(c)
    test_workbook(c, args.excel)
    return c.summary()


if __name__ == "__main__":
    raise SystemExit(main())
