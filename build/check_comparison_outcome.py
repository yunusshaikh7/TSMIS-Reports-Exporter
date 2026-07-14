"""Phase-2 comparison producer contract.

Locks the additive, machine-readable result emitted by ``compare_core.run_compare``:
completion is never implicit, counts are exact and injectively keyed, warnings stay
structured, every output flavor returns the same truth, and error/cancel exits fail
closed.  Workbook serialization remains owned by the existing comparison goldens;
the repeated semantic dump here additionally proves the new metadata is not written
into workbook cells.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_comparison_outcome.py
"""
import sys
import tempfile
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT / "scripts"))

import outcome
from compare_core import CompareSchema, run_compare
from comparison_contract import ComparisonCounts, ComparisonOutcome
from events import Events
from openpyxl import load_workbook


SCHEMA = CompareSchema(
    report_name="Typed outcome",
    header=["Key", "F1", "F2", "Context"],
    side_a="SIDE-A",
    side_b="SIDE-B",
    id_noun="row",
    id_noun_plural="rows",
    context_fields=("Context",),
)

CLEAN = [
    ["k1", "same", "a", "context-a"],
    ["k2", "left", "2", "context-a"],
    ["k3", "last", "3", "context-a"],
]

SIDE_A = [
    ["k1", "same", "a", "context-a"],
    ["k2", "left", "2", "context-a"],
    ["ka", "only-a", "x", "context-a"],
]
SIDE_B = [
    ["k1", "same", "b", "context-b"],
    ["k2", "right", "2", "context-b"],
    ["kb", "only-b", "y", "context-b"],
]


def _assert_typed(result):
    assert isinstance(result.comparison_outcome, ComparisonOutcome), result
    assert isinstance(result.comparison_outcome.counts, ComparisonCounts), result
    assert result.comparison_outcome.status == result.status
    assert result.comparison_outcome.completion == result.completion
    return result.comparison_outcome


def _semantic_dump(path):
    """Workbook cell/type dump: typed return metadata must never leak into XLSX."""
    wb = load_workbook(path, data_only=False, read_only=False)
    try:
        return tuple(
            (ws.title, tuple(
                (cell.coordinate, cell.value, cell.data_type)
                for row in ws.iter_rows()
                for cell in row
                if cell.value is not None
            ))
            for ws in wb.worksheets
        )
    finally:
        wb.close()


def test_clean_and_partial(tmp):
    clean_path = tmp / "clean.xlsx"
    clean = run_compare(SCHEMA, CLEAN, CLEAN, False, clean_path, mode="values")
    typed = _assert_typed(clean)
    assert clean.status == "ok" and clean.completion == outcome.COMPLETE
    assert clean.verdict == "match" and typed.verdict == "match"
    assert typed.warnings == () and typed.failures == ()
    assert clean.skipped_inputs == 0 and clean.failed_inputs == 0
    assert typed.counts == ComparisonCounts(
        known=True,
        paired_rows=3,
        side_a_only_rows=0,
        side_b_only_rows=0,
        differing_rows=0,
        differing_cells=0,
        per_field_counts={"1:F1": 0, "2:F2": 0, "3:Context": 0},
        asserted_cells=6,
        context_cells=3,
    )

    warning = "SIDE-B route_099.xlsx: unreadable (BadZipFile)"
    partial_path = tmp / "partial.xlsx"
    partial = run_compare(
        SCHEMA, CLEAN, CLEAN, False, partial_path, mode="values",
        warnings=[warning],
    )
    typed = _assert_typed(partial)
    assert partial.status == "ok" and partial.completion == outcome.PARTIAL
    assert partial.verdict == "diff" and typed.verdict == "diff"
    assert partial.skipped_inputs == 1 and partial.failed_inputs == 0
    assert typed.warnings == (warning,) and typed.failures == ()
    # Coverage truth and equality truth are independent: the compared rows still
    # match exactly even though the incomplete input set blocks certification.
    assert typed.counts.paired_rows == 3
    assert typed.counts.differing_rows == 0
    assert typed.counts.differing_cells == 0


def test_exact_counts_and_invariants(tmp):
    result = run_compare(
        SCHEMA, SIDE_A, SIDE_B, False, tmp / "diff.xlsx", mode="values")
    typed = _assert_typed(result)
    counts = typed.counts
    assert result.status == "ok" and result.completion == outcome.COMPLETE
    assert result.verdict == "diff" and typed.verdict == "diff"
    assert counts == ComparisonCounts(
        known=True,
        paired_rows=2,
        side_a_only_rows=1,
        side_b_only_rows=1,
        differing_rows=2,
        differing_cells=2,
        per_field_counts={"1:F1": 1, "2:F2": 1, "3:Context": 0},
        asserted_cells=4,
        context_cells=2,
    )
    assert counts.paired_rows + counts.side_a_only_rows + counts.side_b_only_rows == 4
    assert counts.differing_rows <= counts.paired_rows
    assert sum(counts.per_field_counts.values()) == counts.differing_cells
    assert counts.differing_cells <= counts.asserted_cells


def test_structured_input_coverage(tmp):
    result = run_compare(
        SCHEMA, CLEAN, CLEAN, False, tmp / "producer-partial.xlsx",
        mode="values", input_completion=outcome.PARTIAL,
        skipped_inputs=7, failed_inputs=2,
        warnings=("PDF producer skipped seven inputs",),
        failures=("PDF producer failed two inputs",))
    typed = _assert_typed(result)
    assert result.status == "ok" and result.completion == outcome.PARTIAL
    assert result.verdict == "diff"
    assert result.skipped_inputs == 7 and result.failed_inputs == 2
    assert typed.warnings == ("PDF producer skipped seven inputs",)
    assert typed.failures == ("PDF producer failed two inputs",)
    assert typed.counts.paired_rows == 3 and typed.counts.differing_cells == 0


def test_output_mode_parity_and_no_cell_changes(tmp):
    results = {}
    for mode in ("formulas", "values", "both"):
        results[mode] = run_compare(
            SCHEMA, SIDE_A, SIDE_B, False, tmp / f"{mode}.xlsx", mode=mode)
        _assert_typed(results[mode])
    first = results["formulas"].comparison_outcome
    for mode in ("values", "both"):
        current = results[mode].comparison_outcome
        assert current.counts == first.counts
        assert current.completion == first.completion == outcome.COMPLETE
        assert current.verdict == first.verdict == "diff"
        assert current.warnings == first.warnings == ()

    # Two identical values builds have cell-for-cell/type-for-type identical
    # semantics. The result contract is return metadata only.
    repeat = tmp / "repeat.xlsx"
    run_compare(SCHEMA, SIDE_A, SIDE_B, False, repeat, mode="values")
    assert _semantic_dump(tmp / "values.xlsx") == _semantic_dump(repeat)


def test_early_exit_normalization(tmp):
    invalid = run_compare(
        SCHEMA, CLEAN, CLEAN, False, tmp / "invalid.xlsx", mode="not-a-mode")
    typed = _assert_typed(invalid)
    assert invalid.status == "error" and invalid.completion == outcome.FAILED
    assert typed.verdict == "unknown" and not typed.counts.known
    assert typed.failures and "Unknown comparison mode" in typed.failures[0]

    empty = run_compare(
        SCHEMA, [], CLEAN, False, tmp / "empty.xlsx", mode="values")
    typed = _assert_typed(empty)
    assert empty.status == "error" and empty.completion == outcome.FAILED
    assert typed.failures and not typed.counts.known

    cancelled = run_compare(
        SCHEMA, CLEAN, CLEAN, False, tmp / "cancelled.xlsx", mode="values",
        events=Events(is_cancelled=lambda: True),
    )
    typed = _assert_typed(cancelled)
    assert cancelled.status == "cancelled"
    assert cancelled.completion == outcome.CANCELLED
    assert typed.completion == "cancelled" and typed.verdict == "unknown"
    assert typed.failures == () and not typed.counts.known

    existing = tmp / "existing.xlsx"
    existing.write_bytes(b"prior")
    declined = run_compare(
        SCHEMA, CLEAN, CLEAN, False, existing, mode="values",
        confirm_overwrite=lambda _path: False,
    )
    typed = _assert_typed(declined)
    assert declined.status == "cancelled" and declined.completion == outcome.CANCELLED
    assert typed.completion == "cancelled" and existing.read_bytes() == b"prior"


def main():
    with tempfile.TemporaryDirectory(prefix="tsmis_comparison_outcome_") as raw:
        tmp = Path(raw)
        test_clean_and_partial(tmp)
        test_exact_counts_and_invariants(tmp)
        test_structured_input_coverage(tmp)
        test_output_mode_parity_and_no_cell_changes(tmp)
        test_early_exit_normalization(tmp)
    print("OK  COMPARISON-OUTCOME: explicit completion, exact structured counts, "
          "structured warnings, mode parity, normalized early exits, and no "
          "typed metadata in workbook cells.")


if __name__ == "__main__":
    main()
