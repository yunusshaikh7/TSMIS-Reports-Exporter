#!/usr/bin/env python3
"""Independent Stage-8 audit of Highway Sequence product comparison twins.

This checker deliberately imports no product comparison, publication, or
sidecar reader.  It authenticates the frozen witness inputs and the independent
TSMIS source oracle, reads the formula/value workbooks directly with openpyxl,
reconstructs every Comparison cell from the embedded source sheets, and
independently authenticates the schema-v3 generation envelopes and pairing
payloads.

The PDF-vs-Excel leg has two simultaneous contracts which must not be
conflated:

* source semantics (suffix is an asserted field): 60,493 pairs, zero PDF-only,
  one Excel-only;
* the current product's defective glued-suffix identity: 59,946 pairs, 547
  PDF-only, 548 Excel-only (CMP-AUD-199).

Reproducing the second contract is a product-witness PASS, not a claim that the
comparison is source-correct.  The emitted audit JSON says so explicitly.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
from functools import lru_cache
import hashlib
import json
import os
from pathlib import Path
import posixpath
import re
import stat
import xml.etree.ElementTree as ET
import zipfile
import zlib

from openpyxl import load_workbook


VISUAL_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd"
)
DEFAULT_WITNESS_RESULT = (
    VISUAL_ROOT / "phase8_highway_sequence_product_comparisons_r1" / "result.json"
)
SOURCE_ORACLE = VISUAL_ROOT / "phase8_highway_sequence_source_oracle_draft_r1.json"
DEFAULT_OUTPUT = (
    VISUAL_ROOT / "phase8_highway_sequence_product_comparison_parity_r1.json"
)

SOURCE_ORACLE_BINDING = {
    "bytes": 4_008_580,
    "sha256": "2c0997b7d3eb000ac40eddcb5107fa86951ca98825b58394af2b640a5c964b90",
}
INPUT_BINDINGS = {
    "excel": {
        "bytes": 2_424_212,
        "sha256": "cf5905332db3d3eb5a49a87d603f6e36f209cad9a84173b381dace6600168b20",
    },
    "pdf": {
        "bytes": 2_371_547,
        "sha256": "070afe51ea3bf84c9704d0a36a02702b65189941badab6374b03461db8ef6ccc",
    },
    "tsn": {
        "bytes": 2_536_901,
        "sha256": "9dc84c661a9284131baf928767e210a6d708c0a338819fca2b69b907f85dd041",
    },
}

FIELDS = (
    "County", "City", "HG", "FT", "Distance To Next Point", "Description",
)
CONTEXT_FIELDS = frozenset(("City", "HG", "Distance To Next Point"))
ASSERTED_FIELDS = tuple(field for field in FIELDS if field not in CONTEXT_FIELDS)
SOURCE_HEADER = (
    "Route", "County", "PM", "City", "HG", "FT",
    "Distance To Next Point", "Description",
)
STATE_HEADER = "__CMP_E1_STATE_V1_C001_P0000_P0005"
DIFF_MARK = " ≠ "
BUILD_FRESH_RE = re.compile(r"^__CMP_E2_BUILD_FRESH_V1_C\d{3}_B_[A-Z]+$")
HELPER_RE = re.compile(r"^__CMP_E2_KEY_V1_\d+$")
PAYLOAD_RE = re.compile(
    r"\.cmpv3-[0-9a-f]{64}-[0-9]{6}-[0-9a-f]{64}"
    r"(?:-f-(?:0[0-7]|[0-9a-f]{64}-[0-9a-f]{16}))?"
    r"\.comparison-payload\.zlib"
)
PM_RE = re.compile(r"^([CDGHLMNRST]?)(\d{3}\.\d{3})(E?)$")
GENERATION_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-"
    r"[0-9a-f]{4}-[0-9a-f]{12}$"
)

LEG_SPECS = {
    "excel_vs_normalized_tsn": {
        "side_a": "TSMIS", "side_b": "TSN",
        "side_a_rows": 60_494, "side_b_rows": 69_758,
        "paired_rows": 57_072, "side_a_only_rows": 3_422,
        "side_b_only_rows": 12_686,
    },
    "pdf_vs_normalized_tsn": {
        "side_a": "TSMIS (PDF)", "side_b": "TSN",
        "side_a_rows": 60_493, "side_b_rows": 69_758,
        "paired_rows": 57_505, "side_a_only_rows": 2_988,
        "side_b_only_rows": 12_253,
    },
    "pdf_vs_excel": {
        "side_a": "TSMIS (PDF)", "side_b": "TSMIS (Excel)",
        "side_a_rows": 60_493, "side_b_rows": 60_494,
        "paired_rows": 59_946, "side_a_only_rows": 547,
        "side_b_only_rows": 548,
    },
}

SOURCE_SEMANTIC_PDF_VS_EXCEL = {
    "paired_rows": 60_493,
    "side_a_only_rows": 0,
    "side_b_only_rows": 1,
}


class AuditError(RuntimeError):
    """A frozen identity, workbook, or publication contract was violated."""


@dataclass(frozen=True)
class SourceRow:
    """One normalized row copied into a product comparison source sheet."""

    source_index: int
    projection: tuple[object, ...]
    helper: str

    @property
    def route(self) -> str:
        return _trim(self.projection[0])

    @property
    def county(self) -> str:
        return _trim(self.projection[1])

    @property
    def pm(self) -> str:
        return _trim(self.projection[2])

    @property
    def fields(self) -> tuple[str, ...]:
        # SOURCE_HEADER has PM at index 2; Comparison displays every field but PM.
        return tuple(_trim(self.projection[index]) for index in (1, 3, 4, 5, 6, 7))

    @property
    def product_key(self) -> tuple[str, str]:
        return self.route, f"{self.county} {self.pm}".strip()

    @property
    def semantic_key(self) -> tuple[str, ...]:
        match = PM_RE.fullmatch(self.pm)
        if match is not None:
            prefix, base, _suffix = match.groups()
            return self.route, self.county, prefix, base
        if self.pm:
            raise AuditError(
                f"source row {self.source_index + 2}: malformed Highway Sequence PM "
                f"{self.pm!r}"
            )
        return (
            self.route, self.county, "<PMLESS>", self.fields[2], self.fields[3],
            _pmless_kind(self.fields[-1]),
        )


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, allow_nan=False,
        sort_keys=True, separators=(",", ":"),
    ).encode("utf-8")


def _json_line(value: object) -> bytes:
    return _canonical(value) + b"\n"


def _identity(path: Path) -> dict[str, object]:
    path = path.resolve()
    if not path.is_file() or path.is_symlink():
        raise AuditError(f"required regular file is absent or indirect: {path}")
    return {
        "path": str(path), "bytes": path.stat().st_size,
        "sha256": _sha_file(path),
    }


def _strict_json(path: Path, *, maximum: int | None = None) -> dict[str, object]:
    raw = path.read_bytes()
    if maximum is not None and len(raw) > maximum:
        raise AuditError(f"{path.name}: JSON exceeds {maximum:,} bytes")
    try:
        value = json.loads(raw.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise AuditError(f"{path.name}: invalid strict UTF-8 JSON: {exc}") from exc
    if not isinstance(value, dict):
        raise AuditError(f"{path.name}: JSON root is not an object")
    return value


def _trim(value: object) -> str:
    if value is None:
        return ""
    if type(value) is bool:
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(" +", " ", str(value)).strip(" ")


def _pmless_kind(value: object) -> str:
    text = re.sub(r"_x000d_", " ", _trim(value), flags=re.IGNORECASE)
    text = " ".join(text.replace("\r", " ").replace("\n", " ").split()).upper()
    for prefix in ("END OF ROUTE", "CITY END:", "COUNTY END:", "DISTRICT END:"):
        if text.startswith(prefix):
            return prefix
    return text


def _exact_int(value: object, label: str, *, minimum: int = 0) -> int:
    if type(value) is not int or value < minimum:
        raise AuditError(f"{label}: expected integer >= {minimum}, got {value!r}")
    return value


def _bind_source_oracle() -> dict[str, object]:
    observed = _identity(SOURCE_ORACLE)
    if {key: observed[key] for key in ("bytes", "sha256")} != SOURCE_ORACLE_BINDING:
        raise AuditError(f"independent source-oracle identity drift: {observed}")
    document = _strict_json(SOURCE_ORACLE)
    current = document.get("source_parity", {}).get("current_excel_vs_pdf")
    if not isinstance(current, dict):
        raise AuditError("independent source oracle lacks current Excel/PDF parity")
    expected = {
        "left_rows": 60_494, "right_rows": 60_493,
        "paired_rows": 60_493, "left_only_rows": 1, "right_only_rows": 0,
    }
    if {key: current.get(key) for key in expected} != expected:
        raise AuditError("independent source-oracle row contract drift")
    left_only = current.get("left_only")
    if not isinstance(left_only, list) or len(left_only) != 1:
        raise AuditError("independent source oracle lost the single Excel-only row")
    row = left_only[0]
    if (
        row.get("identity") != ["010", "LA", "", "014.814", "occurrence:1"]
        or row.get("ref") != "highway_sequence_route_010.xlsx:row:191"
        or row.get("values", [None] * 9)[-1] != "010/EB ON FR VERMONT"
    ):
        raise AuditError(f"independent source-oracle Excel-only row drift: {row!r}")
    return {
        "identity": observed,
        "audit_complete": document.get("audit_complete"),
        "status": document.get("status"),
        "identity_policy": current.get("identity_policy"),
        "excel_vs_pdf": expected,
        "excel_only": row,
    }


def _formula_tag_census(path: Path) -> dict[str, int]:
    """Count physical formula tags per sheet without trusting openpyxl caches."""
    pattern = re.compile(rb"<(?:[A-Za-z0-9_]+:)?f(?:\s|>)")
    counts: dict[str, int] = {}
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {item.attrib["Id"]: item.attrib["Target"] for item in relationships}
        ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        for sheet in workbook.findall("main:sheets/main:sheet", ns):
            label = sheet.attrib["name"]
            target = rel_targets[sheet.attrib[f"{{{ns['rel']}}}id"]].replace("\\", "/")
            member = (
                target.lstrip("/") if target.startswith("/")
                else posixpath.normpath(posixpath.join("xl", target))
            )
            count = 0
            tail = b""
            with archive.open(member) as handle:
                for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                    payload = tail + chunk
                    count += sum(match.end() > len(tail) for match in pattern.finditer(payload))
                    tail = payload[-96:]
            counts[label] = count
    return counts


def _cell_values(cells, width: int) -> tuple[object, ...]:
    cells = tuple(cells)
    values = tuple(cell.value for cell in cells)
    return values + (None,) * max(0, width - len(values))


def _reject_errors(cells, label: str) -> None:
    for cell in cells:
        if cell.data_type == "e":
            raise AuditError(f"{label}!{cell.coordinate}: error cell {cell.value!r}")


def _read_source_sheet(worksheet, expected_rows: int) -> tuple[list[SourceRow], dict[str, object]]:
    physical = iter(worksheet.iter_rows())
    header_cells = tuple(next(physical, ()))
    header = tuple(cell.value for cell in header_cells)
    prefix = ("Comparison row", *SOURCE_HEADER, "Key (helper)")
    if len(header) <= len(prefix) or header[:len(prefix)] != prefix:
        raise AuditError(f"{worksheet.title}: source header drift: {header!r}")
    freshness_headers = header[len(prefix):]
    if not freshness_headers or any(
        not isinstance(value, str) or BUILD_FRESH_RE.fullmatch(value) is None
        for value in freshness_headers
    ) or len(set(freshness_headers)) != len(freshness_headers):
        raise AuditError(f"{worksheet.title}: build-freshness header drift")
    if any(cell.data_type in {"f", "e"} for cell in header_cells):
        raise AuditError(f"{worksheet.title}: formula/error in source header")

    rows: list[SourceRow] = []
    helpers = set()
    footer_count = 0
    formula_cells = 0
    width = len(header)
    for physical_row, cells in enumerate(physical, 2):
        cells = tuple(cells)
        _reject_errors(cells, worksheet.title)
        values = _cell_values(cells, width)
        route = values[1]
        if route is None or _trim(route) == "":
            if not any(value is not None for value in values):
                continue
            if any(value is not None for value in values[:len(prefix)]):
                raise AuditError(f"{worksheet.title}: malformed sparse row {physical_row}")
            tail_cells = cells[len(prefix):]
            if len(tail_cells) != len(freshness_headers) or any(
                cell.data_type != "f" for cell in tail_cells
            ):
                raise AuditError(f"{worksheet.title}: appended-row footer drift")
            formula_cells += len(tail_cells)
            footer_count += 1
            continue
        if len(cells) != width:
            raise AuditError(f"{worksheet.title}: data row {physical_row} width drift")
        if cells[0].data_type != "f":
            raise AuditError(f"{worksheet.title}: missing Comparison backlink at {physical_row}")
        if any(cell.data_type in {"f", "e"} for cell in cells[1:len(prefix)]):
            raise AuditError(f"{worksheet.title}: formula/error entered source projection")
        if any(cell.data_type != "f" for cell in cells[len(prefix):]):
            raise AuditError(f"{worksheet.title}: build-freshness formula omission")
        helper = values[9]
        if not isinstance(helper, str) or HELPER_RE.fullmatch(helper) is None or helper in helpers:
            raise AuditError(f"{worksheet.title}: helper-token drift at row {physical_row}")
        helpers.add(helper)
        rows.append(SourceRow(len(rows), tuple(values[1:9]), helper))
        formula_cells += 1 + len(freshness_headers)
    if len(rows) != expected_rows or footer_count != 1:
        raise AuditError(
            f"{worksheet.title}: {len(rows):,} rows/{footer_count} footers, expected "
            f"{expected_rows:,}/1"
        )
    digest = _sha_bytes(_canonical([
        [row.source_index, *row.projection, row.helper] for row in rows
    ]))
    return rows, {
        "rows": len(rows), "columns": width, "footer_count": footer_count,
        "freshness_columns": len(freshness_headers),
        "formula_cells": formula_cells,
        "ordered_projection_sha256": digest,
        "helper_tokens_unique": len(helpers) == len(rows),
    }


def _inspect_snapshot(worksheet, source: list[SourceRow]) -> dict[str, object]:
    if worksheet.sheet_state != "veryHidden":
        raise AuditError(f"{worksheet.title}: snapshot is not veryHidden")
    physical = iter(worksheet.iter_rows())
    header = _cell_values(next(physical, ()), 10)
    expected_header = ("Source row", *SOURCE_HEADER, "Key (helper)")
    if header != expected_header:
        raise AuditError(f"{worksheet.title}: snapshot header drift")
    observed = []
    for ordinal, cells in enumerate(physical, 1):
        cells = tuple(cells)
        values = _cell_values(cells, 10)
        if not any(value is not None for value in values):
            continue
        if len(cells) != 10 or any(cell.data_type in {"f", "e"} for cell in cells):
            raise AuditError(f"{worksheet.title}: snapshot row shape drift")
        if ordinal > len(source):
            raise AuditError(f"{worksheet.title}: extra snapshot row")
        expected = (ordinal, *source[ordinal - 1].projection, source[ordinal - 1].helper)
        if values != expected:
            raise AuditError(f"{worksheet.title}: snapshot mismatch at row {ordinal + 1}")
        observed.append(values)
    if len(observed) != len(source):
        raise AuditError(f"{worksheet.title}: snapshot row census drift")
    return {
        "rows": len(observed), "columns": 10,
        "ordered_snapshot_sha256": _sha_bytes(_canonical(observed)),
        "source_projection_exact": True,
    }


def _sheet_digest(worksheet, *, require_no_formulas: bool) -> dict[str, object]:
    rows = []
    for cells in worksheet.iter_rows():
        cells = tuple(cells)
        _reject_errors(cells, worksheet.title)
        if require_no_formulas and any(cell.data_type == "f" for cell in cells):
            raise AuditError(f"{worksheet.title}: unexpected formula")
        values = tuple(cell.value for cell in cells)
        if any(value is not None for value in values):
            rows.append(values)
    if not rows:
        raise AuditError(f"{worksheet.title}: unexpectedly empty sheet")
    return {
        "rows": len(rows), "ordered_cells_sha256": _sha_bytes(_canonical(rows)),
    }


_LINK_RE = re.compile(
    r'^=HYPERLINK\("#(?P<sheet>(?:\'(?:[^\']|\'\')+\'|[^!]+))!'
    r'(?P<row>\d+):(?P=row)",(?P<label>\d+)\)$'
)


def _link_source_index(cell, side: str, label: str) -> int:
    if cell.data_type != "f" or not isinstance(cell.value, str):
        raise AuditError(f"{label}: expected a HYPERLINK formula")
    match = _LINK_RE.fullmatch(cell.value)
    if match is None:
        raise AuditError(f"{label}: malformed HYPERLINK {cell.value!r}")
    sheet = match.group("sheet")
    if sheet.startswith("'"):
        sheet = sheet[1:-1].replace("''", "'")
    row = int(match.group("row"))
    if sheet != side or row != int(match.group("label")) or row < 2:
        raise AuditError(f"{label}: HYPERLINK target drift {cell.value!r}")
    return row - 2


def _expected_field_projection(
    left: SourceRow | None, right: SourceRow | None,
) -> tuple[tuple[str, ...], str]:
    if left is None and right is None:
        raise AuditError("Comparison row has neither source")
    if left is None:
        return right.fields, "U" * len(FIELDS)
    if right is None:
        return left.fields, "U" * len(FIELDS)
    displays = []
    states = []
    for field, value_a, value_b in zip(FIELDS, left.fields, right.fields, strict=True):
        if field in CONTEXT_FIELDS:
            states.append("N")
            displays.append(value_a if value_a else value_b)
        elif value_a == value_b:
            states.append("E")
            displays.append(value_a)
        else:
            states.append("D")
            displays.append(
                f"{value_a or '(blank)'}{DIFF_MARK}{value_b or '(blank)'}"
            )
    return tuple(displays), "".join(states)


def _comparison_header(side_a: str, side_b: str) -> tuple[str, ...]:
    return (
        "Route", "PM", "#", f"{side_a} Row", f"{side_b} Row",
        "Status", "Diffs", *FIELDS, STATE_HEADER,
    )


def _only_header(present: str, missing: str) -> tuple[str, ...]:
    return (
        "Route", "PM", "#", f"{present} Row", f"Missing from {missing}",
        *FIELDS,
    )


def _inspect_value_comparison(
    worksheet, side_a: str, side_b: str,
    rows_a: list[SourceRow], rows_b: list[SourceRow],
) -> dict[str, object]:
    physical = iter(worksheet.iter_rows())
    expected_header = _comparison_header(side_a, side_b)
    header_cells = tuple(next(physical, ()))
    if _cell_values(header_cells, len(expected_header)) != expected_header:
        raise AuditError(f"{worksheet.title}: Comparison header drift")
    if any(cell.data_type in {"f", "e"} for cell in header_cells):
        raise AuditError("Comparison header contains formula/error")

    row_ids = []
    one_a = []
    one_b = []
    paired_source_indices: set[tuple[int, int]] = set()
    used_a: Counter[int] = Counter()
    used_b: Counter[int] = Counter()
    status_counts: Counter[str] = Counter()
    per_field: Counter[str] = Counter()
    differing_rows = differing_cells = asserted_cells = context_cells = 0
    union_rows = 0
    for physical_row, cells in enumerate(physical, 2):
        cells = tuple(cells)
        values = _cell_values(cells, len(expected_header))
        if not any(value is not None for value in values):
            continue
        if len(cells) != len(expected_header):
            raise AuditError(f"Comparison row {physical_row}: width drift")
        _reject_errors(cells, "Comparison")
        if any(cell.data_type == "f" for cell in (*cells[:3], *cells[5:])):
            raise AuditError(f"Comparison row {physical_row}: value cell is a formula")

        route = _trim(values[0])
        visible_key = _trim(values[1])
        occurrence = _exact_int(values[2], f"Comparison row {physical_row} occurrence", minimum=1)
        index_a = (
            _link_source_index(cells[3], side_a, f"Comparison row {physical_row} side A")
            if values[3] is not None else None
        )
        index_b = (
            _link_source_index(cells[4], side_b, f"Comparison row {physical_row} side B")
            if values[4] is not None else None
        )
        if index_a is not None and index_a >= len(rows_a):
            raise AuditError(f"Comparison row {physical_row}: side-A row is out of range")
        if index_b is not None and index_b >= len(rows_b):
            raise AuditError(f"Comparison row {physical_row}: side-B row is out of range")
        left = rows_a[index_a] if index_a is not None else None
        right = rows_b[index_b] if index_b is not None else None
        if left is None and right is None:
            raise AuditError(f"Comparison row {physical_row}: neither side is present")
        for source in (left, right):
            if source is not None and source.product_key != (route, visible_key):
                raise AuditError(
                    f"Comparison row {physical_row}: displayed identity does not bind source"
                )
        status = (
            "Both" if left is not None and right is not None
            else f"{side_a} only" if left is not None else f"{side_b} only"
        )
        if values[5] != status:
            raise AuditError(f"Comparison row {physical_row}: status drift")
        displays, state = _expected_field_projection(left, right)
        observed_displays = tuple(_trim(value) for value in values[7:13])
        if observed_displays != displays or values[13] != state:
            raise AuditError(
                f"Comparison row {physical_row}: display/state differs from embedded sources"
            )
        expected_diffs = state.count("D") if status == "Both" else None
        if values[6] != expected_diffs:
            raise AuditError(f"Comparison row {physical_row}: Diffs/state disagreement")
        if status == "Both":
            if set(state) - {"E", "D", "N"}:
                raise AuditError(f"Comparison row {physical_row}: invalid matched state")
            differing_rows += int(expected_diffs > 0)
            differing_cells += expected_diffs
            for field, code in zip(FIELDS, state, strict=True):
                per_field[field] += int(code == "D")
                if code == "N":
                    context_cells += 1
                else:
                    asserted_cells += 1
            paired_source_indices.add((index_a, index_b))
        elif state != "U" * len(FIELDS):
            raise AuditError(f"Comparison row {physical_row}: one-sided state drift")

        if index_a is not None:
            used_a[index_a] += 1
        if index_b is not None:
            used_b[index_b] += 1
        inventory = (
            route, visible_key, occurrence,
            index_a if left is not None else index_b, *displays,
        )
        if status == f"{side_a} only":
            one_a.append(inventory)
        elif status == f"{side_b} only":
            one_b.append(inventory)
        status_counts[status] += 1
        row_ids.append((route, visible_key, occurrence))
        union_rows += 1

    if used_a != Counter(range(len(rows_a))) or used_b != Counter(range(len(rows_b))):
        raise AuditError("Comparison source-row coverage is not exactly once per side")
    counts = {
        "known": True,
        "paired_rows": status_counts["Both"],
        "side_a_only_rows": status_counts[f"{side_a} only"],
        "side_b_only_rows": status_counts[f"{side_b} only"],
        "differing_rows": differing_rows,
        "differing_cells": differing_cells,
        "per_field_counts": {
            field: per_field[field] for field in FIELDS if per_field[field]
        },
        "asserted_cells": asserted_cells,
        "context_cells": context_cells,
    }
    return {
        "counts": counts, "union_rows": union_rows,
        "row_ids": row_ids,
        "side_a_only_inventory": one_a,
        "side_b_only_inventory": one_b,
        "paired_source_indices": paired_source_indices,
        "paired_source_index_sha256": _sha_bytes(
            _canonical(sorted([list(pair) for pair in paired_source_indices]))
        ),
        "comparison_rows_reconstructed_from_sources": True,
    }


def _inspect_only_values(
    worksheet, *, present: str, missing: str,
    expected: list[tuple[object, ...]], present_rows: list[SourceRow],
    missing_rows: list[SourceRow],
) -> dict[str, object]:
    physical = iter(worksheet.iter_rows())
    expected_header = _only_header(present, missing)
    header_cells = tuple(next(physical, ()))
    if _cell_values(header_cells, len(expected_header)) != expected_header:
        raise AuditError(f"{worksheet.title}: one-sided header drift")
    observed = []
    missing_routes = {row.route for row in missing_rows}
    for physical_row, cells in enumerate(physical, 2):
        cells = tuple(cells)
        values = _cell_values(cells, len(expected_header))
        if not any(value is not None for value in values):
            continue
        if len(cells) != len(expected_header):
            raise AuditError(f"{worksheet.title}: row width drift")
        _reject_errors(cells, worksheet.title)
        if cells[3].data_type != "f" or any(
            cell.data_type == "f" for cell in (*cells[:3], *cells[4:])
        ):
            raise AuditError(f"{worksheet.title}: values/formula role drift")
        index = _link_source_index(cells[3], present, f"{worksheet.title} row {physical_row}")
        if index >= len(present_rows):
            raise AuditError(f"{worksheet.title}: source row out of range")
        source = present_rows[index]
        route, visible_key = source.product_key
        occurrence = _exact_int(values[2], f"{worksheet.title} occurrence", minimum=1)
        displays = tuple(_trim(value) for value in values[5:])
        if (_trim(values[0]), _trim(values[1]), displays) != (
            route, visible_key, source.fields,
        ):
            raise AuditError(f"{worksheet.title}: source projection drift")
        reason = "this location only" if route in missing_routes else "entire route"
        if values[4] != reason:
            raise AuditError(f"{worksheet.title}: missing-scope classification drift")
        observed.append((route, visible_key, occurrence, index, *displays))
    if observed != expected:
        raise AuditError(f"{worksheet.title}: inventory differs from Comparison sheet")
    return {
        "rows": len(observed), "inventory_exact": True,
        "ordered_inventory_sha256": _sha_bytes(_canonical(observed)),
    }


def _read_routes_values(worksheet, expected_routes: set[str]) -> list[str]:
    expected_header = (
        "Route", "Status", None, None, "Locations compared",
        "Matched locations", "Locations w/ differences", "Differing cells",
    )
    physical = iter(worksheet.iter_rows())
    header_cells = tuple(next(physical, ()))
    header = _cell_values(header_cells, 8)
    # Side-specific row labels occupy C/D; bind all stable labels independently.
    if (
        len(header) != 8 or header[0] != expected_header[0]
        or header[1] != expected_header[1] or header[4:] != expected_header[4:]
    ):
        raise AuditError("Routes header drift")
    routes = []
    for cells in physical:
        cells = tuple(cells)
        values = _cell_values(cells, 8)
        if not any(value is not None for value in values):
            continue
        if len(cells) != 8 or any(cell.data_type in {"f", "e"} for cell in cells):
            raise AuditError("values Routes row formula/error drift")
        routes.append(_trim(values[0]))
    if len(routes) != len(set(routes)) or set(routes) != expected_routes:
        raise AuditError("Routes sheet route universe drift")
    return routes


def _inspect_formula_comparison(worksheet, expected_row_ids: list[tuple[str, str, int]],
                                side_a: str, side_b: str) -> dict[str, object]:
    expected_header = _comparison_header(side_a, side_b)
    physical = iter(worksheet.iter_rows())
    if _cell_values(next(physical, ()), len(expected_header)) != expected_header:
        raise AuditError("formula Comparison header drift")
    rows = 0
    for cells in physical:
        cells = tuple(cells)
        values = _cell_values(cells, len(expected_header))
        if not any(value is not None for value in values):
            continue
        if rows >= len(expected_row_ids) or len(cells) != len(expected_header):
            raise AuditError("formula Comparison row census/width drift")
        _reject_errors(cells, "formula Comparison")
        observed_id = (_trim(values[0]), _trim(values[1]), values[2])
        if observed_id != expected_row_ids[rows]:
            raise AuditError(f"formula Comparison identity drift at row {rows + 2}")
        if any(cell.data_type in {"f", "e"} for cell in cells[:3]) or any(
            cell.data_type != "f" for cell in cells[3:]
        ):
            raise AuditError(f"formula Comparison formula-role drift at row {rows + 2}")
        rows += 1
    if rows != len(expected_row_ids):
        raise AuditError("formula Comparison omitted rows")
    return {"rows": rows, "formulas_per_row": len(expected_header) - 3}


def _inspect_formula_only(worksheet, expected: list[tuple[object, ...]],
                          present: str, missing: str) -> dict[str, object]:
    expected_header = _only_header(present, missing)
    physical = iter(worksheet.iter_rows())
    if _cell_values(next(physical, ()), len(expected_header)) != expected_header:
        raise AuditError(f"{worksheet.title}: formula one-sided header drift")
    rows = 0
    for cells in physical:
        cells = tuple(cells)
        values = _cell_values(cells, len(expected_header))
        if not any(value is not None for value in values):
            continue
        if rows >= len(expected) or len(cells) != len(expected_header):
            raise AuditError(f"{worksheet.title}: formula row census drift")
        if (_trim(values[0]), _trim(values[1]), values[2]) != expected[rows][:3]:
            raise AuditError(f"{worksheet.title}: formula identity drift")
        if any(cell.data_type in {"f", "e"} for cell in cells[:3]) or any(
            cell.data_type != "f" for cell in cells[3:]
        ):
            raise AuditError(f"{worksheet.title}: formula-role drift")
        rows += 1
    if rows != len(expected):
        raise AuditError(f"{worksheet.title}: formula rows omitted")
    return {"rows": rows, "formulas_per_row": len(expected_header) - 3}


def _inspect_formula_routes(worksheet, expected_routes: list[str]) -> dict[str, object]:
    physical = iter(worksheet.iter_rows())
    header = tuple(cell.value for cell in next(physical, ()))
    if len(header) != 8 or header[0:2] != ("Route", "Status"):
        raise AuditError("formula Routes header drift")
    routes = []
    for cells in physical:
        cells = tuple(cells)
        values = _cell_values(cells, 8)
        if not any(value is not None for value in values):
            continue
        if len(cells) != 8 or cells[0].data_type in {"f", "e"} or any(
            cell.data_type != "f" for cell in cells[1:]
        ):
            raise AuditError("formula Routes formula-role drift")
        routes.append(_trim(values[0]))
    if routes != expected_routes:
        raise AuditError("formula/value Routes order drift")
    return {"rows": len(routes), "formulas_per_row": 7}


def _multiset_alignment(
    left: list[SourceRow], right: list[SourceRow], *, semantic: bool,
) -> dict[str, int]:
    left_counts = Counter(
        row.semantic_key if semantic else row.product_key for row in left
    )
    right_counts = Counter(
        row.semantic_key if semantic else row.product_key for row in right
    )
    paired = sum(
        min(left_counts[key], right_counts[key])
        for key in set(left_counts) | set(right_counts)
    )
    return {
        "paired_rows": paired,
        "side_a_only_rows": len(left) - paired,
        "side_b_only_rows": len(right) - paired,
    }


def _pair_cost(left: SourceRow, right: SourceRow) -> int:
    return sum(
        value_a != value_b
        for field, value_a, value_b in zip(FIELDS, left.fields, right.fields, strict=True)
        if field not in CONTEXT_FIELDS
    )


def _optimal_assignment(cost: list[list[int]]) -> tuple[int, list[int], str]:
    rows = len(cost)
    columns = len(cost[0]) if cost else 0
    if not rows or not columns or any(len(row) != columns for row in cost):
        raise AuditError("pairing trace supplied an empty/ragged cost matrix")
    if max(rows, columns) > 12:
        raise AuditError(
            f"independent trace solver bound exceeded: {rows}x{columns}"
        )
    smaller_side = "a" if rows <= columns else "b"
    oriented = cost if smaller_side == "a" else [
        [cost[a_index][b_index] for a_index in range(rows)]
        for b_index in range(columns)
    ]
    small = len(oriented)
    large = len(oriented[0])

    @lru_cache(maxsize=None)
    def solve(index: int, used: int) -> tuple[int, tuple[int, ...]]:
        if index == small:
            return 0, ()
        best: tuple[int, tuple[int, ...]] | None = None
        for candidate in range(large):
            if used & (1 << candidate):
                continue
            tail_cost, tail_vector = solve(index + 1, used | (1 << candidate))
            value = oriented[index][candidate] + tail_cost, (candidate, *tail_vector)
            if best is None or value < best:
                best = value
        if best is None:
            raise AuditError("independent assignment solver found no candidate")
        return best

    total, vector = solve(0, 0)
    return total, list(vector), smaller_side


def _normalize_counts(value: object, label: str) -> dict[str, object]:
    if not isinstance(value, dict):
        raise AuditError(f"{label}: counts are absent")
    expected_fields = {
        "known", "paired_rows", "side_a_only_rows", "side_b_only_rows",
        "differing_rows", "differing_cells", "per_field_counts",
        "asserted_cells", "context_cells",
    }
    if set(value) != expected_fields:
        raise AuditError(f"{label}: count field universe drift: {set(value)!r}")
    per_field = value.get("per_field_counts")
    if not isinstance(per_field, dict):
        raise AuditError(f"{label}: per-field counts are absent")
    normalized_fields: dict[str, int] = {}
    for raw_field, raw_count in per_field.items():
        field = str(raw_field).split(":", 1)[-1]
        count = _exact_int(raw_count, f"{label} {field}")
        if not count:
            continue
        if field in normalized_fields:
            raise AuditError(f"{label}: duplicate normalized field {field!r}")
        normalized_fields[field] = count
    result = dict(value)
    if result["known"] is not True:
        raise AuditError(f"{label}: counts are not known")
    for field in expected_fields - {"known", "per_field_counts"}:
        result[field] = _exact_int(result[field], f"{label} {field}")
    result["per_field_counts"] = normalized_fields
    return result


def _validate_pairing_trace(
    trace_value: object, rows_a: list[SourceRow], rows_b: list[SourceRow],
    workbook_pairs: set[tuple[int, int]], label: str,
) -> dict[str, object]:
    if not isinstance(trace_value, list):
        raise AuditError(f"{label}: pairing trace is not an array")
    groups_a: dict[tuple[str, str], list[int]] = defaultdict(list)
    groups_b: dict[tuple[str, str], list[int]] = defaultdict(list)
    for index, row in enumerate(rows_a):
        groups_a[row.product_key].append(index)
    for index, row in enumerate(rows_b):
        groups_b[row.product_key].append(index)
    workbook_pairs_by_key: dict[tuple[str, str], set[tuple[int, int]]] = defaultdict(set)
    for pair in workbook_pairs:
        key_a = rows_a[pair[0]].product_key
        key_b = rows_b[pair[1]].product_key
        if key_a != key_b:
            raise AuditError(f"{label}: workbook paired different product keys")
        workbook_pairs_by_key[key_a].add(pair)
    duplicate_keys = {
        key for key in set(groups_a) & set(groups_b)
        if max(len(groups_a[key]), len(groups_b[key])) > 1
    }
    if len(trace_value) != len(duplicate_keys):
        raise AuditError(
            f"{label}: pairing trace group census {len(trace_value)} != "
            f"independent {len(duplicate_keys)}"
        )
    seen_keys = set()
    semantic_trace = []
    expected_fields = {
        "key_components", "side_a_size", "side_b_size", "matrix_cells",
        "side_a_indices", "side_b_indices", "smaller_side",
        "assignment_vector", "pairs", "total_cost", "positional_cost",
        "algorithm", "exact", "quality",
    }
    for ordinal, item in enumerate(trace_value):
        prefix = f"{label} trace {ordinal}"
        if not isinstance(item, dict) or set(item) != expected_fields:
            raise AuditError(f"{prefix}: field universe drift")
        components = item["key_components"]
        if (
            not isinstance(components, list) or len(components) != 2
            or not all(isinstance(value, str) for value in components)
        ):
            raise AuditError(f"{prefix}: key-components shape drift")
        key = tuple(components)
        if key not in duplicate_keys or key in seen_keys:
            raise AuditError(f"{prefix}: missing/duplicate/non-duplicate key {key!r}")
        seen_keys.add(key)
        indices_a = item["side_a_indices"]
        indices_b = item["side_b_indices"]
        if indices_a != groups_a[key] or indices_b != groups_b[key]:
            raise AuditError(f"{prefix}: source-index inventory drift")
        size_a = _exact_int(item["side_a_size"], f"{prefix} side A", minimum=1)
        size_b = _exact_int(item["side_b_size"], f"{prefix} side B", minimum=1)
        if size_a != len(indices_a) or size_b != len(indices_b):
            raise AuditError(f"{prefix}: declared group sizes drift")
        matrix_cells = _exact_int(item["matrix_cells"], f"{prefix} matrix", minimum=1)
        if matrix_cells != size_a * size_b or matrix_cells > 100_000:
            raise AuditError(f"{prefix}: matrix/cap contract drift")
        cost = [
            [_pair_cost(rows_a[a_index], rows_b[b_index]) for b_index in indices_b]
            for a_index in indices_a
        ]
        optimal_cost, optimal_vector, smaller_side = _optimal_assignment(cost)
        if (
            item["algorithm"] != "rectangular-hungarian-lex-v1"
            or item["exact"] is not True or item["quality"] != "exact"
            or item["smaller_side"] != smaller_side
            or item["assignment_vector"] != optimal_vector
            or item["total_cost"] != optimal_cost
        ):
            raise AuditError(f"{prefix}: exact/lexicographic assignment drift")
        positional = sum(cost[index][index] for index in range(min(size_a, size_b)))
        if item["positional_cost"] != positional or optimal_cost > positional:
            raise AuditError(f"{prefix}: positional-cost contract drift")

        reconstructed = []
        for small_index, large_index in enumerate(optimal_vector):
            if smaller_side == "a":
                reconstructed.append((indices_a[small_index], indices_b[large_index]))
            else:
                reconstructed.append((indices_a[large_index], indices_b[small_index]))
        pair_items = item["pairs"]
        if not isinstance(pair_items, list) or len(pair_items) != len(reconstructed):
            raise AuditError(f"{prefix}: pair-record census drift")
        observed_pairs = []
        observed_cost = 0
        for pair_item, expected_pair in zip(pair_items, reconstructed, strict=True):
            if not isinstance(pair_item, dict) or set(pair_item) != {
                "side_a_index", "side_b_index", "cost",
            }:
                raise AuditError(f"{prefix}: pair-record shape drift")
            pair = (pair_item["side_a_index"], pair_item["side_b_index"])
            if pair != expected_pair:
                raise AuditError(f"{prefix}: pair/vector disagreement")
            actual_cost = _pair_cost(rows_a[pair[0]], rows_b[pair[1]])
            if pair_item["cost"] != actual_cost:
                raise AuditError(f"{prefix}: per-pair cost drift")
            observed_cost += actual_cost
            observed_pairs.append(pair)
        if observed_cost != optimal_cost:
            raise AuditError(f"{prefix}: pair/total cost disagreement")
        actual_group_pairs = workbook_pairs_by_key.get(key, set())
        if actual_group_pairs != set(observed_pairs):
            raise AuditError(f"{prefix}: sidecar/workbook pair disagreement")
        semantic_trace.append({
            "key_components": components,
            "side_a_size": size_a, "side_b_size": size_b,
            "matrix_cells": matrix_cells, "smaller_side": smaller_side,
            "assignment_vector": optimal_vector, "total_cost": optimal_cost,
            "source_pairs": [list(pair) for pair in sorted(observed_pairs)],
        })
    if seen_keys != duplicate_keys:
        raise AuditError(f"{label}: pairing trace omitted duplicate groups")
    return {
        "groups": len(semantic_trace),
        "max_group_size": max(
            (max(item["side_a_size"], item["side_b_size"]) for item in semantic_trace),
            default=0,
        ),
        "semantic_trace_sha256": _sha_bytes(_canonical(semantic_trace)),
        "wire_trace_sha256": _sha_bytes(_canonical(trace_value)),
        "all_assignments_independently_exact_and_lexicographic": True,
        "all_trace_pairs_match_workbook": True,
    }


def _publication_member(
    member: object, expected_path: Path, flavor: str, bound: dict[str, object], label: str,
) -> dict[str, object]:
    if not isinstance(member, dict):
        raise AuditError(f"{label}: publication member is not an object")
    stat_result = expected_path.stat()
    expected_role = "canonical" if flavor == "values" else "best_effort"
    expected = {
        "path": str(expected_path.resolve()),
        "relative_path": expected_path.name,
        "size": bound["bytes"], "sha256": bound["sha256"],
        "mtime_ns": stat_result.st_mtime_ns, "flavor": flavor,
        "commit_role": expected_role,
        "canonical_path_at_write": str(expected_path.resolve()).casefold(),
    }
    for key, value in expected.items():
        observed = member.get(key)
        if key == "path":
            try:
                observed = str(Path(str(observed)).resolve())
            except (OSError, RuntimeError):
                pass
        if observed != value:
            raise AuditError(f"{label}: publication member {key} drift")
    return expected


def _inflate_payload(
    manifest: dict[str, object], parent: Path, label: str,
) -> tuple[dict[str, object], list[dict[str, object]], set[Path]]:
    if set(manifest) != {
        "schema_version", "encoding", "decoded_size", "decoded_sha256",
        "binding_sha256", "chunks",
    } or (manifest.get("schema_version"), manifest.get("encoding")) != (
        1, "canonical-json-zlib-chunks-v1",
    ):
        raise AuditError(f"{label}: payload manifest schema drift")
    chunks = manifest.get("chunks")
    if not isinstance(chunks, list) or not chunks:
        raise AuditError(f"{label}: empty payload chunk manifest")
    decoded_parts = []
    details = []
    referenced: set[Path] = set()
    for ordinal, descriptor in enumerate(chunks):
        if not isinstance(descriptor, dict) or set(descriptor) != {
            "decoded_size", "relative_path", "sha256", "size",
        }:
            raise AuditError(f"{label}: payload descriptor shape drift")
        relative = descriptor["relative_path"]
        if (
            not isinstance(relative, str) or PAYLOAD_RE.fullmatch(relative) is None
            or Path(relative).name != relative
        ):
            raise AuditError(f"{label}: unsafe payload path {relative!r}")
        path = parent / relative
        resolved = path.resolve()
        if resolved in referenced or resolved.parent != parent.resolve():
            raise AuditError(f"{label}: duplicate/escaping payload path")
        referenced.add(resolved)
        identity = _identity(path)
        if identity["bytes"] != descriptor["size"] or identity["sha256"] != descriptor["sha256"]:
            raise AuditError(f"{label}: payload chunk identity drift")
        if identity["bytes"] > 67_108_864:
            raise AuditError(f"{label}: payload chunk exceeds audit bound")
        raw = path.read_bytes()
        inflater = zlib.decompressobj()
        try:
            decoded = inflater.decompress(raw) + inflater.flush()
        except zlib.error as exc:
            raise AuditError(f"{label}: payload cannot inflate: {exc}") from exc
        if (
            not inflater.eof or inflater.unused_data or inflater.unconsumed_tail
            or len(decoded) != descriptor["decoded_size"]
        ):
            raise AuditError(f"{label}: payload chunk framing drift")
        decoded_parts.append(decoded)
        details.append({
            "ordinal": ordinal, "path": str(resolved),
            "bytes": len(raw), "sha256": _sha_bytes(raw),
            "decoded_bytes": len(decoded),
        })
    decoded = b"".join(decoded_parts)
    if (
        len(decoded) != manifest["decoded_size"]
        or _sha_bytes(decoded) != manifest["decoded_sha256"]
        or len(decoded) > 67_108_864
    ):
        raise AuditError(f"{label}: aggregate payload identity drift")
    try:
        persisted = json.loads(decoded.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise AuditError(f"{label}: payload JSON drift: {exc}") from exc
    if not isinstance(persisted, dict) or _canonical(persisted) != decoded:
        raise AuditError(f"{label}: payload is not canonical JSON")
    return persisted, details, referenced


def _inspect_publication(
    label: str, formulas_path: Path, values_path: Path,
    leg_witness: dict[str, object], workbook_counts: dict[str, object],
    rows_a: list[SourceRow], rows_b: list[SourceRow],
    workbook_pairs: set[tuple[int, int]], identities: dict[str, dict[str, object]],
) -> tuple[dict[str, object], set[Path]]:
    paths = {"formulas": formulas_path, "values": values_path}
    envelopes: dict[str, dict[str, object]] = {}
    sidecars = {}
    generation_members = None
    for flavor, workbook in paths.items():
        sidecar_path = Path(str(workbook) + ".outcome.json")
        sentinel = Path(str(sidecar_path) + ".tmp")
        if sentinel.exists():
            raise AuditError(f"{label}: publication sentinel remains: {sentinel}")
        sidecar_identity = _identity(sidecar_path)
        raw = sidecar_path.read_bytes()
        if len(raw) > 1_048_576:
            raise AuditError(f"{label}: sidecar exceeds audit bound")
        envelope = _strict_json(sidecar_path, maximum=1_048_576)
        if _canonical(envelope) != raw:
            raise AuditError(f"{label}: sidecar is not canonical JSON")
        if set(envelope) != {
            "schema_version", "comparison_schema_version", "record_type",
            "completion", "skipped_inputs", "failed_inputs", "built_at_mtime",
            "artifact_generation", "self_member", "comparison_payload",
        }:
            raise AuditError(f"{label}: sidecar field universe drift")
        if (
            envelope["schema_version"], envelope["comparison_schema_version"],
            envelope["record_type"], envelope["completion"],
            envelope["skipped_inputs"], envelope["failed_inputs"],
        ) != (1, 3, "comparison", "complete", 0, 0):
            raise AuditError(f"{label}: sidecar terminal-state drift")
        if envelope["built_at_mtime"] != workbook.stat().st_mtime:
            raise AuditError(f"{label}: sidecar build-time mtime drift")
        generation = envelope["artifact_generation"]
        if not isinstance(generation, dict) or set(generation) != {
            "generation_id", "completion", "publication_state", "requested_mode",
            "members", "content_digests", "producer_versions",
        }:
            raise AuditError(f"{label}: generation field universe drift")
        if (
            generation["completion"], generation["publication_state"],
            generation["requested_mode"],
        ) != ("complete", "committed", "both"):
            raise AuditError(f"{label}: generation state drift")
        if not isinstance(generation["generation_id"], str) or GENERATION_RE.fullmatch(
            generation["generation_id"]
        ) is None:
            raise AuditError(f"{label}: malformed generation ID")
        if not isinstance(generation["producer_versions"], dict):
            raise AuditError(f"{label}: producer version map drift")
        members = generation["members"]
        if not isinstance(members, list) or len(members) != 2:
            raise AuditError(f"{label}: twin member census drift")
        by_flavor = {
            member.get("flavor"): member for member in members if isinstance(member, dict)
        }
        if set(by_flavor) != set(paths):
            raise AuditError(f"{label}: twin member flavor drift")
        inspected = {
            item_flavor: _publication_member(
                by_flavor[item_flavor], item_path, item_flavor,
                identities[item_flavor], f"{label} {flavor} envelope",
            )
            for item_flavor, item_path in paths.items()
        }
        if envelope["self_member"] != by_flavor[flavor]:
            raise AuditError(f"{label}: sidecar self-member drift")
        expected_digests = {
            item_flavor: identities[item_flavor]["sha256"] for item_flavor in paths
        }
        if generation["content_digests"] != expected_digests:
            raise AuditError(f"{label}: generation digest map drift")
        if generation_members is None:
            generation_members = inspected
        envelopes[flavor] = envelope
        sidecars[flavor] = sidecar_identity

    formula_envelope = envelopes["formulas"]
    value_envelope = envelopes["values"]
    generation = formula_envelope["artifact_generation"]
    manifest = formula_envelope["comparison_payload"]
    if generation != value_envelope["artifact_generation"]:
        raise AuditError(f"{label}: formula/value generations diverge")
    if manifest != value_envelope["comparison_payload"]:
        raise AuditError(f"{label}: formula/value payload manifests diverge")
    persisted, chunks, referenced = _inflate_payload(
        manifest, formulas_path.parent, label,
    )
    binding = _sha_bytes(_canonical({
        "decoded_sha256": manifest["decoded_sha256"],
        "completion": "complete", "skipped_inputs": 0, "failed_inputs": 0,
        "artifact_generation": generation,
    }))
    if manifest["binding_sha256"] != binding:
        raise AuditError(f"{label}: payload/generation binding drift")
    expected_payload_fields = {
        "status", "completion", "verdict", "counts", "warnings", "failures",
        "source_identities", "pairing_trace", "duplicate_group_count",
        "pairing_quality", "capped_group_diagnostics", "coverage_diagnostics",
    }
    if set(persisted) != expected_payload_fields:
        raise AuditError(f"{label}: persisted outcome field universe drift")
    counts = _normalize_counts(persisted["counts"], f"{label} persisted")
    expected_verdict = (
        "match" if not counts["differing_cells"]
        and not counts["side_a_only_rows"] and not counts["side_b_only_rows"]
        else "diff"
    )
    if (
        persisted["status"], persisted["completion"], persisted["verdict"],
        persisted["pairing_quality"],
    ) != ("ok", "complete", expected_verdict, "exact"):
        raise AuditError(f"{label}: persisted terminal/pairing state drift")
    if counts != workbook_counts:
        raise AuditError(f"{label}: persisted/workbook count disagreement")
    for key in (
        "warnings", "failures", "source_identities", "capped_group_diagnostics",
        "coverage_diagnostics",
    ):
        if persisted[key] != []:
            raise AuditError(f"{label}: persisted {key} is not empty")
    trace = _validate_pairing_trace(
        persisted["pairing_trace"], rows_a, rows_b, workbook_pairs, label,
    )
    if persisted["duplicate_group_count"] != trace["groups"]:
        raise AuditError(f"{label}: duplicate-group count/trace disagreement")

    returned = leg_witness.get("result")
    if not isinstance(returned, dict):
        raise AuditError(f"{label}: witness returned-result object is absent")
    if (
        returned.get("status"), returned.get("completion"), returned.get("verdict"),
        returned.get("skipped_inputs"), returned.get("failed_inputs"),
        returned.get("pairing_quality"), returned.get("pairing_trace_count"),
    ) != (
        "ok", "complete", expected_verdict, 0, 0, "exact", trace["groups"],
    ):
        raise AuditError(f"{label}: returned typed terminal state drift")
    if returned.get("warnings") != [] or returned.get("failures") != []:
        raise AuditError(f"{label}: returned warnings/failures are not empty")
    returned_counts = _normalize_counts(returned.get("counts"), f"{label} returned")
    if returned_counts != workbook_counts:
        raise AuditError(f"{label}: returned/workbook count disagreement")
    returned_generation = returned.get("artifact_generation")
    if not isinstance(returned_generation, dict) or set(returned_generation) != {
        "completion", "publication_state", "requested_mode", "members",
    }:
        raise AuditError(f"{label}: returned generation shape drift")
    if (
        returned_generation["completion"], returned_generation["publication_state"],
        returned_generation["requested_mode"],
    ) != ("complete", "committed", "both"):
        raise AuditError(f"{label}: returned generation state drift")
    returned_members = returned_generation["members"]
    if not isinstance(returned_members, list) or len(returned_members) != 2:
        raise AuditError(f"{label}: returned member census drift")
    returned_by_flavor = {
        item.get("flavor"): item for item in returned_members if isinstance(item, dict)
    }
    persisted_by_flavor = {
        item["flavor"]: item for item in generation["members"]
    }
    for flavor, path in paths.items():
        persisted_member = persisted_by_flavor[flavor]
        expected_return = {
            "flavor": flavor, "commit_role": persisted_member["commit_role"],
            "path": str(path.resolve()), "bytes": persisted_member["size"],
            "sha256": persisted_member["sha256"],
        }
        observed_return = returned_by_flavor.get(flavor)
        if isinstance(observed_return, dict) and "path" in observed_return:
            observed_return = dict(observed_return)
            observed_return["path"] = str(Path(str(observed_return["path"])).resolve())
        if observed_return != expected_return:
            raise AuditError(f"{label}: returned/persisted member disagreement")
    return ({
        "generation_id": generation["generation_id"],
        "generation_binding_sha256": binding,
        "sidecars": sidecars,
        "generation_members": generation_members,
        "twin_generation_and_payload_exact": True,
        "payload": {
            "decoded_bytes": manifest["decoded_size"],
            "decoded_sha256": manifest["decoded_sha256"], "chunks": chunks,
        },
        "persisted_counts": counts,
        "pairing_trace": trace,
        "returned_persisted_workbook_counts_exact": True,
    }, referenced)


def _expected_sheet_names(side_a: str, side_b: str) -> list[str]:
    return [
        "Summary", "Spot Check", "Comparison", "Routes",
        f"Only in {side_a}", f"Only in {side_b}", side_a, side_b,
        "Notes", "__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B",
    ]


def _verify_formula_census(
    formula_census: dict[str, int], value_census: dict[str, int],
    expected_sheets: list[str], comparison: dict[str, object],
    formula_sources: dict[str, dict[str, object]],
    value_sources: dict[str, dict[str, object]],
    side_a: str, side_b: str, route_count: int,
) -> dict[str, object]:
    if list(formula_census) != expected_sheets or list(value_census) != expected_sheets:
        raise AuditError("formula/value XML sheet universe drift")
    counts = comparison["counts"]
    union = comparison["union_rows"]
    expected_formula = {
        "Comparison": (len(_comparison_header(side_a, side_b)) - 3) * union,
        "Routes": 7 * route_count,
        f"Only in {side_a}": (len(_only_header(side_a, side_b)) - 3)
        * counts["side_a_only_rows"],
        f"Only in {side_b}": (len(_only_header(side_b, side_a)) - 3)
        * counts["side_b_only_rows"],
        side_a: formula_sources[side_a]["formula_cells"],
        side_b: formula_sources[side_b]["formula_cells"],
        "__CMP_E2_SNAPSHOT_A": 0, "__CMP_E2_SNAPSHOT_B": 0, "Notes": 0,
    }
    expected_values = {
        "Comparison": 2 * counts["paired_rows"]
        + counts["side_a_only_rows"] + counts["side_b_only_rows"],
        "Routes": 0,
        f"Only in {side_a}": counts["side_a_only_rows"],
        f"Only in {side_b}": counts["side_b_only_rows"],
        side_a: value_sources[side_a]["formula_cells"],
        side_b: value_sources[side_b]["formula_cells"],
        "__CMP_E2_SNAPSHOT_A": 0, "__CMP_E2_SNAPSHOT_B": 0, "Notes": 0,
    }
    for sheet, expected in expected_formula.items():
        if formula_census[sheet] != expected:
            raise AuditError(f"formula workbook {sheet!r} formula census drift")
    for sheet, expected in expected_values.items():
        if value_census[sheet] != expected:
            raise AuditError(f"values workbook {sheet!r} formula census drift")
    if (
        formula_census["Summary"] <= value_census["Summary"]
        or value_census["Summary"] <= 0
        or formula_census["Spot Check"] != value_census["Spot Check"]
        or value_census["Spot Check"] <= 0
        or sum(formula_census.values()) <= sum(value_census.values())
    ):
        raise AuditError("formula/value fixed-sheet richness contract drift")
    return {
        "formulas": formula_census, "values": value_census,
        "critical_sheet_counts_exact": True,
        "formula_flavor_richer": True,
    }


def _inspect_leg_workbooks(
    label: str, spec: dict[str, object], formulas_path: Path, values_path: Path,
) -> dict[str, object]:
    side_a = str(spec["side_a"])
    side_b = str(spec["side_b"])
    expected_sheets = _expected_sheet_names(side_a, side_b)
    formula_census = _formula_tag_census(formulas_path)
    value_census = _formula_tag_census(values_path)

    values = load_workbook(
        values_path, read_only=True, data_only=False, keep_links=False,
    )
    try:
        if values.sheetnames != expected_sheets:
            raise AuditError(f"{label}: values sheet universe/order drift")
        rows_a, source_a = _read_source_sheet(values[side_a], int(spec["side_a_rows"]))
        rows_b, source_b = _read_source_sheet(values[side_b], int(spec["side_b_rows"]))
        snapshot_a = _inspect_snapshot(values["__CMP_E2_SNAPSHOT_A"], rows_a)
        snapshot_b = _inspect_snapshot(values["__CMP_E2_SNAPSHOT_B"], rows_b)
        comparison = _inspect_value_comparison(
            values["Comparison"], side_a, side_b, rows_a, rows_b,
        )
        only_a = _inspect_only_values(
            values[f"Only in {side_a}"], present=side_a, missing=side_b,
            expected=comparison["side_a_only_inventory"],
            present_rows=rows_a, missing_rows=rows_b,
        )
        only_b = _inspect_only_values(
            values[f"Only in {side_b}"], present=side_b, missing=side_a,
            expected=comparison["side_b_only_inventory"],
            present_rows=rows_b, missing_rows=rows_a,
        )
        expected_routes = {row.route for row in (*rows_a, *rows_b)}
        routes = _read_routes_values(values["Routes"], expected_routes)
        notes = _sheet_digest(values["Notes"], require_no_formulas=True)
        value_sources = {side_a: source_a, side_b: source_b}
    finally:
        values.close()

    fixed_counts = {
        key: int(spec[key]) for key in (
            "paired_rows", "side_a_only_rows", "side_b_only_rows",
        )
    }
    if {key: comparison["counts"][key] for key in fixed_counts} != fixed_counts:
        raise AuditError(
            f"{label}: product row identity counts drift from independently frozen census"
        )
    product_alignment = _multiset_alignment(rows_a, rows_b, semantic=False)
    if product_alignment != fixed_counts:
        raise AuditError(f"{label}: embedded sources do not explain product row counts")

    formulas = load_workbook(
        formulas_path, read_only=True, data_only=False, keep_links=False,
    )
    try:
        if formulas.sheetnames != expected_sheets:
            raise AuditError(f"{label}: formulas sheet universe/order drift")
        formula_rows_a, formula_source_a = _read_source_sheet(
            formulas[side_a], int(spec["side_a_rows"]),
        )
        formula_rows_b, formula_source_b = _read_source_sheet(
            formulas[side_b], int(spec["side_b_rows"]),
        )
        if formula_rows_a != rows_a or formula_rows_b != rows_b:
            raise AuditError(f"{label}: formula/value embedded source sheets diverge")
        formula_snapshot_a = _inspect_snapshot(
            formulas["__CMP_E2_SNAPSHOT_A"], formula_rows_a,
        )
        formula_snapshot_b = _inspect_snapshot(
            formulas["__CMP_E2_SNAPSHOT_B"], formula_rows_b,
        )
        formula_comparison = _inspect_formula_comparison(
            formulas["Comparison"], comparison["row_ids"], side_a, side_b,
        )
        formula_only_a = _inspect_formula_only(
            formulas[f"Only in {side_a}"], comparison["side_a_only_inventory"],
            side_a, side_b,
        )
        formula_only_b = _inspect_formula_only(
            formulas[f"Only in {side_b}"], comparison["side_b_only_inventory"],
            side_b, side_a,
        )
        formula_routes = _inspect_formula_routes(formulas["Routes"], routes)
        formula_notes = _sheet_digest(formulas["Notes"], require_no_formulas=True)
        if formula_notes != notes:
            raise AuditError(f"{label}: formula/value Notes sheets diverge")
        formula_sources = {side_a: formula_source_a, side_b: formula_source_b}
    finally:
        formulas.close()

    formula_contract = _verify_formula_census(
        formula_census, value_census, expected_sheets, comparison,
        formula_sources, value_sources, side_a, side_b, len(routes),
    )
    semantic_alignment = None
    if label == "pdf_vs_excel":
        semantic_alignment = _multiset_alignment(rows_a, rows_b, semantic=True)
        if semantic_alignment != SOURCE_SEMANTIC_PDF_VS_EXCEL:
            raise AuditError(
                f"{label}: embedded sources violate independent source-semantic pairing"
            )
        semantic_counts_a = Counter(row.semantic_key for row in rows_a)
        semantic_counts_b = Counter(row.semantic_key for row in rows_b)
        semantic_only_b = [
            row for row in rows_b
            if semantic_counts_b[row.semantic_key] > semantic_counts_a[row.semantic_key]
        ]
        # There is exactly one multiplicity excess; bind its source-visible claims.
        if not any(
            row.route == "010" and row.county == "LA" and row.pm == "014.814"
            and row.fields[-1] == "EB ON FR VERMONT"
            for row in semantic_only_b
        ):
            raise AuditError("pdf_vs_excel: expected route-010 Excel-only row is absent")

    return {
        "sheet_universe": expected_sheets,
        "source_sheets": {
            "values": value_sources, "formulas": formula_sources,
            "formula_value_rows_exact": True,
        },
        "snapshots": {
            "values": {side_a: snapshot_a, side_b: snapshot_b},
            "formulas": {side_a: formula_snapshot_a, side_b: formula_snapshot_b},
            "formula_value_snapshot_exact": (
                snapshot_a == formula_snapshot_a and snapshot_b == formula_snapshot_b
            ),
        },
        "comparison": comparison,
        "only_in": {side_a: only_a, side_b: only_b},
        "formula_shell": {
            "Comparison": formula_comparison,
            f"Only in {side_a}": formula_only_a,
            f"Only in {side_b}": formula_only_b,
            "Routes": formula_routes,
        },
        "formula_contract": formula_contract,
        "product_key_alignment": product_alignment,
        "source_semantic_alignment": semantic_alignment,
        "rows_a": rows_a, "rows_b": rows_b,
    }


def _bind_output_paths(
    label: str, leg: dict[str, object], spec: dict[str, object],
) -> tuple[Path, Path, dict[str, dict[str, object]]]:
    outputs = leg.get("outputs")
    if not isinstance(outputs, dict) or set(outputs) != {"formulas", "values"}:
        raise AuditError(f"{label}: witness output identity map drift")
    paths = {}
    identities = {}
    for flavor, declared in outputs.items():
        if not isinstance(declared, dict) or set(declared) != {"path", "bytes", "sha256"}:
            raise AuditError(f"{label}: {flavor} output identity shape drift")
        path = Path(str(declared["path"])).resolve()
        if path.suffix.casefold() != ".xlsx" or ".tmp" in path.name.casefold():
            raise AuditError(f"{label}: unsafe {flavor} output basename: {path.name!r}")
        observed = _identity(path)
        declared_normalized = dict(declared)
        declared_normalized["path"] = str(path)
        if observed != declared_normalized:
            raise AuditError(f"{label}: {flavor} output identity drift")
        paths[flavor] = path
        identities[flavor] = observed
    if paths["formulas"].parent != paths["values"].parent:
        raise AuditError(f"{label}: formula/value twins are in different directories")
    expected_values = paths["formulas"].with_name(
        f"{paths['formulas'].stem} (values){paths['formulas'].suffix}"
    )
    if paths["values"] != expected_values:
        raise AuditError(f"{label}: formula/value twin naming contract drift")
    return paths["formulas"], paths["values"], identities


def _validate_loaded_product_code(value: object) -> dict[str, object]:
    if not isinstance(value, dict):
        raise AuditError("loaded-product-code manifest shape drift")
    legacy = set(value) == {"files", "canonical_json_sha256", "members"}
    split = set(value) == {
        "schema", "files", "bytes", "canonical_members_sha256", "members",
    }
    if not legacy and not split:
        raise AuditError("loaded-product-code manifest field universe drift")
    members = value["members"]
    if not isinstance(members, list) or value["files"] != len(members) or not members:
        raise AuditError("loaded-product-code member census drift")
    expected_digest = (
        value["canonical_json_sha256"] if legacy
        else value["canonical_members_sha256"]
    )
    observed_digest = _sha_bytes(_json_line(members) if legacy else _canonical(members))
    if observed_digest != expected_digest:
        raise AuditError("loaded-product-code canonical digest drift")
    if split and (
        value["schema"] != "phase8-loaded-product-code-manifest/v1"
        or value["bytes"] != sum(item.get("bytes", -1) for item in members if isinstance(item, dict))
    ):
        raise AuditError("loaded-product-code schema/byte total drift")
    repo_root = Path(__file__).resolve().parent.parent
    details = []
    for item in members:
        expected_member_fields = (
            {"module", "relative_path", "bytes", "sha256"} if legacy
            else {"relative_path", "modules", "bytes", "sha256"}
        )
        if not isinstance(item, dict) or set(item) != expected_member_fields:
            raise AuditError("loaded-product-code member shape drift")
        if split and (
            not isinstance(item["modules"], list)
            or item["modules"] != sorted(item["modules"])
            or not all(isinstance(name, str) for name in item["modules"])
        ):
            raise AuditError("loaded-product-code module aliases drift")
        relative = Path(str(item["relative_path"]))
        path = (repo_root / "scripts" / relative).resolve()
        scripts_root = (repo_root / "scripts").resolve()
        try:
            path.relative_to(scripts_root)
        except ValueError as exc:
            raise AuditError("loaded-product-code path escapes scripts/") from exc
        observed = _identity(path)
        if (observed["bytes"], observed["sha256"]) != (item["bytes"], item["sha256"]):
            raise AuditError(f"loaded product code changed after witness: {relative}")
        details.append({
            "modules": item["modules"] if split else [item["module"]],
            "relative_path": item["relative_path"], **observed,
        })
    return {
        "schema": value.get("schema", "legacy-monolithic/v1"),
        "files": len(details), "canonical_members_sha256": expected_digest,
        "all_current": True, "members": details,
    }


def _validate_tree_manifest(
    root: Path, manifest: object, *, ignored_names: frozenset[str], label: str,
) -> dict[str, object]:
    if not isinstance(manifest, dict):
        raise AuditError(f"{label}: tree manifest shape drift")
    legacy = set(manifest) == {
        "files", "bytes", "canonical_json_sha256", "members",
    }
    split = set(manifest) == {
        "schema", "scope", "excluded_names", "files", "bytes",
        "canonical_members_sha256", "members",
    }
    if not legacy and not split:
        raise AuditError(f"{label}: tree manifest field universe drift")
    if split:
        if (
            manifest["schema"] != "phase8-local-artifact-manifest/v1"
            or manifest["scope"]
            != "all flat files present before artifact-manifest.json/result.json"
            or manifest["excluded_names"] != sorted(manifest["excluded_names"])
            or not all(isinstance(item, str) for item in manifest["excluded_names"])
        ):
            raise AuditError(f"{label}: split-leg manifest schema drift")
        ignored_names = frozenset(manifest["excluded_names"])
    declared = manifest["members"]
    if not isinstance(declared, list) or manifest["files"] != len(declared):
        raise AuditError(f"{label}: tree manifest member census drift")
    expected_digest = (
        manifest["canonical_json_sha256"] if legacy
        else manifest["canonical_members_sha256"]
    )
    observed_digest = _sha_bytes(_json_line(declared) if legacy else _canonical(declared))
    if observed_digest != expected_digest:
        raise AuditError(f"{label}: tree manifest canonical digest drift")
    actual = []
    iterator = (
        sorted(root.iterdir(), key=lambda item: item.name.casefold()) if split
        else sorted(
            (item for item in root.rglob("*") if item.is_file()),
            key=lambda item: item.relative_to(root).as_posix(),
        )
    )
    for path in iterator:
        if path.relative_to(root).as_posix() in ignored_names:
            continue
        try:
            facts = path.stat(follow_symlinks=False)
        except OSError as exc:
            raise AuditError(f"{label}: cannot stat manifest member {path}") from exc
        reparse = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
        if (
            not stat.S_ISREG(facts.st_mode)
            or bool(getattr(facts, "st_file_attributes", 0) & reparse)
        ):
            raise AuditError(f"{label}: indirect/non-flat manifest member {path}")
        entry = {
            "relative_path": path.relative_to(root).as_posix(),
            "bytes": path.stat().st_size, "sha256": _sha_file(path),
        }
        if split:
            entry["path"] = str(path.resolve())
            entry = {
                "relative_path": entry["relative_path"], "path": entry["path"],
                "bytes": entry["bytes"], "sha256": entry["sha256"],
            }
        actual.append(entry)
    if actual != declared:
        raise AuditError(f"{label}: current tree differs from declared manifest")
    if manifest["bytes"] != sum(item["bytes"] for item in actual):
        raise AuditError(f"{label}: tree byte total drift")
    return {
        "root": str(root.resolve()), "files": len(actual),
        "bytes": manifest["bytes"],
        "canonical_members_sha256": expected_digest,
        "exact": True,
    }


def _collect_tree_manifests(
    document: dict[str, object], result_path: Path,
) -> list[dict[str, object]]:
    """Accept the original one-root witness and composed per-leg witnesses.

    The comparison run was later split into independently committed legs to
    avoid one long wrapper timeout.  A composed witness may therefore put an
    ``artifact_manifest_before_result`` on each leg (optionally with
    ``artifact_root``), or expose ``artifact_manifests`` descriptors.  All
    discovered manifests are mandatory and exact; no output identity is inferred
    from a manifest.
    """
    found = []
    top = document.get("artifact_manifest_before_result")
    if top is not None:
        found.append({
            "label": "combined", "root": result_path.parent,
            "manifest": top, "ignored": frozenset((result_path.name,)),
        })
    descriptors = document.get("artifact_manifests")
    if descriptors is not None:
        if not isinstance(descriptors, dict):
            raise AuditError("artifact_manifests must be an object")
        for label, descriptor in descriptors.items():
            if not isinstance(descriptor, dict):
                raise AuditError(f"artifact manifest {label}: descriptor drift")
            manifest = descriptor.get("manifest", descriptor.get("artifact_manifest_before_result"))
            root = descriptor.get("root", descriptor.get("artifact_root"))
            if manifest is None or not isinstance(root, str):
                raise AuditError(f"artifact manifest {label}: root/manifest absent")
            ignored = descriptor.get("ignored_relative_paths", ["result.json"])
            if not isinstance(ignored, list) or not all(isinstance(item, str) for item in ignored):
                raise AuditError(f"artifact manifest {label}: ignored paths drift")
            found.append({
                "label": str(label), "root": Path(root).resolve(),
                "manifest": manifest, "ignored": frozenset(ignored),
            })
    legs = document.get("legs")
    if isinstance(legs, dict):
        for label, leg in legs.items():
            if not isinstance(leg, dict) or leg.get("artifact_manifest_before_result") is None:
                continue
            outputs = leg.get("outputs")
            formula_path = None
            if isinstance(outputs, dict) and isinstance(outputs.get("formulas"), dict):
                formula_path = outputs["formulas"].get("path")
            root = leg.get("artifact_root")
            root_path = (
                Path(str(root)).resolve() if root is not None
                else Path(str(formula_path)).resolve().parent if formula_path is not None
                else None
            )
            if root_path is None:
                raise AuditError(f"{label}: cannot locate per-leg manifest root")
            found.append({
                "label": str(label), "root": root_path,
                "manifest": leg["artifact_manifest_before_result"],
                "ignored": frozenset(("result.json",)),
            })
    # Avoid silently checking the same descriptor twice when a composer retained
    # both a top-level descriptor map and per-leg convenience copies.
    unique = {}
    for item in found:
        key = (str(item["root"]), _sha_bytes(_canonical(item["manifest"])))
        unique[key] = item
    if not unique:
        raise AuditError("witness declares no artifact tree manifest")
    return list(unique.values())


def _bind_witness_inputs(document: dict[str, object]) -> dict[str, object]:
    inputs = document.get("inputs")
    inputs_after = document.get("inputs_after")
    if not isinstance(inputs, dict) or set(inputs) != set(INPUT_BINDINGS):
        raise AuditError("witness input binding universe drift")
    if inputs_after != inputs:
        raise AuditError("witness inputs changed during product run")
    bound = {}
    for label, expected in INPUT_BINDINGS.items():
        declared = inputs[label]
        if not isinstance(declared, dict) or set(declared) != {"path", "bytes", "sha256"}:
            raise AuditError(f"{label}: witness input identity shape drift")
        path = Path(str(declared["path"])).resolve()
        observed = _identity(path)
        normalized = dict(declared)
        normalized["path"] = str(path)
        if observed != normalized or {
            key: observed[key] for key in ("bytes", "sha256")
        } != expected:
            raise AuditError(f"{label}: frozen witness input identity drift")
        bound[label] = observed
    return bound


def _declared_identity(value: object, label: str) -> tuple[Path, dict[str, object]]:
    if not isinstance(value, dict) or set(value) != {"path", "bytes", "sha256"}:
        raise AuditError(f"{label}: declared file identity shape drift")
    path = Path(str(value["path"])).resolve()
    observed = _identity(path)
    normalized = dict(value)
    normalized["path"] = str(path)
    if observed != normalized:
        raise AuditError(f"{label}: declared file identity drift")
    return path, observed


def _validate_split_leg_record(
    label: str, record: dict[str, object], declared_result: object | None,
) -> dict[str, object]:
    if record.get("leg") != label:
        raise AuditError(f"{label}: split-leg record identifies {record.get('leg')!r}")
    invariants = record.get("invariants")
    if not isinstance(invariants, dict) or not invariants or not all(
        value is True for value in invariants.values()
    ):
        raise AuditError(f"{label}: split-leg runner invariant failed")
    root = Path(str(record.get("output_root", ""))).resolve()
    if not root.is_dir():
        raise AuditError(f"{label}: split-leg output root is absent")
    result_path = root / "result.json"
    result_identity = _identity(result_path)
    if _strict_json(result_path) != record:
        raise AuditError(f"{label}: embedded split-leg record differs from result.json")
    if declared_result is not None:
        declared_path, declared = _declared_identity(
            declared_result, f"{label} composed result reference",
        )
        if declared_path != result_path or declared != result_identity:
            raise AuditError(f"{label}: composed result reference points elsewhere")

    artifact_path, artifact_identity = _declared_identity(
        record.get("artifact_manifest"), f"{label} artifact-manifest.json",
    )
    product_path, product_identity = _declared_identity(
        record.get("product_code_manifest"), f"{label} product-code-manifest.json",
    )
    if artifact_path != root / "artifact-manifest.json":
        raise AuditError(f"{label}: artifact-manifest path drift")
    if product_path != root / "product-code-manifest.json":
        raise AuditError(f"{label}: product-code-manifest path drift")
    if artifact_path.read_bytes() != _json_line(record.get("artifact_manifest_before_result")):
        raise AuditError(f"{label}: artifact-manifest file/content disagreement")
    if product_path.read_bytes() != _json_line(record.get("loaded_product_code")):
        raise AuditError(f"{label}: product-code-manifest file/content disagreement")
    residue = record.get("residue_gate")
    if not isinstance(residue, dict) or residue.get("transient_residue") != []:
        raise AuditError(f"{label}: split-leg residue gate is absent/red")
    return {
        "result": result_identity,
        "artifact_manifest": artifact_identity,
        "product_code_manifest": product_identity,
        "output_root": str(root),
    }


def _extract_leg_documents(
    document: dict[str, object], composed_path: Path,
) -> tuple[dict[str, dict[str, object]], dict[str, object], bool]:
    """Return raw per-leg records or the legacy monolithic leg map."""
    raw_map = document.get("leg_results", document.get("legs"))
    if not isinstance(raw_map, dict) or set(raw_map) != set(LEG_SPECS):
        raise AuditError("composed witness must identify exactly the three legs")
    split_records: dict[str, dict[str, object]] = {}
    split_identities = {}
    is_split = False
    for label, item in raw_map.items():
        declared = None
        record = None
        if isinstance(item, dict) and item.get("leg") == label:
            record = item
            is_split = True
        elif isinstance(item, dict):
            for key in ("result", "witness_result", "result_identity"):
                candidate = item.get(key)
                if isinstance(candidate, dict) and set(candidate) == {
                    "path", "bytes", "sha256",
                }:
                    declared = candidate
                    path, _identity_record = _declared_identity(
                        candidate, f"{label} composed result",
                    )
                    record = _strict_json(path)
                    is_split = True
                    break
            if record is None and isinstance(item.get("witness"), dict):
                record = item["witness"]
                is_split = True
        if record is None:
            # The original monolithic witness stores a compact leg object here.
            continue
        if not isinstance(record, dict):
            raise AuditError(f"{label}: referenced leg result is not an object")
        split_identities[label] = _validate_split_leg_record(label, record, declared)
        split_records[label] = record
    if is_split:
        if set(split_records) != set(LEG_SPECS):
            raise AuditError("composed split witness mixes raw and monolithic leg forms")
        return split_records, split_identities, True
    return {
        label: item for label, item in raw_map.items() if isinstance(item, dict)
    }, {"combined_result": _identity(composed_path)}, False


def _load_composed_witness(path: Path) -> tuple[dict[str, object], dict[str, object]]:
    if not path.is_file():
        raise AuditError(
            f"product comparison witness is incomplete or absent: {path}. "
            "Pass --result for the composed per-leg witness."
        )
    identity = _identity(path)
    document = _strict_json(path)
    legs = document.get("leg_results", document.get("legs"))
    if not isinstance(legs, dict) or set(legs) != set(LEG_SPECS):
        raise AuditError("composed witness must contain exactly the three Highway Sequence legs")
    invariants = document.get("invariants")
    if invariants is not None and (
        not isinstance(invariants, dict) or not all(value is True for value in invariants.values())
    ):
        raise AuditError("witness runner reported a failed invariant")
    return document, identity


def _compose_leg_result_arguments(values: list[str]) -> tuple[dict[str, object], dict[str, object]]:
    references = {}
    identities = {}
    for raw in values:
        if "=" not in raw:
            raise AuditError(
                f"--leg-result must be LEG=PATH, got {raw!r}"
            )
        label, raw_path = raw.split("=", 1)
        if label not in LEG_SPECS or label in references or not raw_path:
            raise AuditError(f"invalid/duplicate --leg-result label: {label!r}")
        path = Path(raw_path).expanduser().resolve()
        identity = _identity(path)
        references[label] = {"result": identity}
        identities[label] = identity
    if set(references) != set(LEG_SPECS):
        raise AuditError(
            "--leg-result must name each of: " + ", ".join(LEG_SPECS)
        )
    document = {
        "audit": "in-memory composition of three immutable split-leg witnesses",
        "leg_results": references,
    }
    composition = {
        "mode": "three immutable --leg-result identities",
        "canonical_reference_sha256": _sha_bytes(_canonical(references)),
        "members": identities,
    }
    return document, composition


def _public_leg_result(leg: dict[str, object]) -> dict[str, object]:
    """Drop in-memory row objects before serializing the audit artifact."""
    result = dict(leg)
    result.pop("rows_a", None)
    result.pop("rows_b", None)
    comparison = dict(result["comparison"])
    comparison.pop("row_ids", None)
    comparison.pop("side_a_only_inventory", None)
    comparison.pop("side_b_only_inventory", None)
    comparison.pop("paired_source_indices", None)
    result["comparison"] = comparison
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    source = parser.add_mutually_exclusive_group()
    source.add_argument(
        "--result", type=Path, default=None,
        help="composed witness JSON (legacy monolithic and split-leg maps supported)",
    )
    source.add_argument(
        "--leg-result", action="append", default=[], metavar="LEG=PATH",
        help="repeat exactly three times to compose immutable split-leg result.json files",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args(argv)

    if args.leg_result:
        witness, witness_identity = _compose_leg_result_arguments(args.leg_result)
        composed_path = args.output.resolve()
    else:
        composed_path = (args.result or DEFAULT_WITNESS_RESULT).resolve()
        witness, witness_identity = _load_composed_witness(composed_path)
    source_oracle = _bind_source_oracle()
    leg_documents, leg_witness_identities, split_witness = _extract_leg_documents(
        witness, composed_path,
    )
    if split_witness:
        input_maps = {
            label: _bind_witness_inputs(record)
            for label, record in leg_documents.items()
        }
        inputs = input_maps[next(iter(LEG_SPECS))]
        if any(item != inputs for item in input_maps.values()):
            raise AuditError("split-leg witnesses disagree on frozen inputs")
        product_code = {
            label: _validate_loaded_product_code(record.get("loaded_product_code"))
            for label, record in leg_documents.items()
        }
        manifest_descriptors = []
        for label, record in leg_documents.items():
            leg_result_path = Path(str(record["output_root"])) / "result.json"
            manifest_descriptors.extend(
                _collect_tree_manifests(record, leg_result_path)
            )
    else:
        inputs = _bind_witness_inputs(witness)
        product_code = {
            "combined": _validate_loaded_product_code(
                witness.get("loaded_product_code")
            )
        }
        manifest_descriptors = _collect_tree_manifests(
            witness, composed_path,
        )
    tree_manifests = [
        _validate_tree_manifest(
            item["root"], item["manifest"], ignored_names=item["ignored"],
            label=f"artifact tree {item['label']}",
        )
        for item in manifest_descriptors
    ]

    audited_legs = {}
    referenced_payloads: set[Path] = set()
    generation_ids = set()
    for label, spec in LEG_SPECS.items():
        leg_witness = leg_documents[label]
        if not isinstance(leg_witness, dict):
            raise AuditError(f"{label}: witness leg is not an object")
        formulas_path, values_path, identities = _bind_output_paths(
            label, leg_witness, spec,
        )
        workbook = _inspect_leg_workbooks(
            label, spec, formulas_path, values_path,
        )
        publication, referenced = _inspect_publication(
            label, formulas_path, values_path, leg_witness,
            workbook["comparison"]["counts"], workbook["rows_a"], workbook["rows_b"],
            workbook["comparison"]["paired_source_indices"], identities,
        )
        if publication["generation_id"] in generation_ids:
            raise AuditError("comparison legs reused a generation ID")
        generation_ids.add(publication["generation_id"])
        referenced_payloads.update(referenced)
        audited_legs[label] = {
            "outputs": identities,
            "workbook": _public_leg_result(workbook),
            "publication": publication,
        }

    declared_payloads = set()
    for manifest in tree_manifests:
        root = Path(manifest["root"])
        declared_payloads.update(
            path.resolve() for path in root.rglob("*.comparison-payload.zlib")
            if path.is_file()
        )
    if declared_payloads != referenced_payloads:
        raise AuditError(
            "artifact manifests contain an unreferenced payload or omit a referenced payload"
        )

    same_source = audited_legs["pdf_vs_excel"]["workbook"]
    if (
        same_source["product_key_alignment"] != {
            "paired_rows": 59_946, "side_a_only_rows": 547,
            "side_b_only_rows": 548,
        }
        or same_source["source_semantic_alignment"] != SOURCE_SEMANTIC_PDF_VS_EXCEL
    ):
        raise AuditError("PDF-vs-Excel source/product identity boundary drift")

    result = {
        "audit": "Stage 8 Highway Sequence product comparison twin audit",
        "status": "pass_with_expected_product_defect",
        "acceptance_artifact": False,
        "reason_not_acceptance": (
            "The current PDF-vs-Excel product uses PM suffix as identity and therefore "
            "reproduces CMP-AUD-199; source semantics require 60,493/0/1, not "
            "59,946/547/548. This artifact authenticates the observed product output."
        ),
        "witness": {
            "composed": witness_identity,
            "leg_results": leg_witness_identities,
            "split_leg_witness": split_witness,
        },
        "source_oracle": source_oracle,
        "inputs": inputs,
        "loaded_product_code": product_code,
        "artifact_tree_manifests": tree_manifests,
        "legs": audited_legs,
        "invariants": {
            "three_complete_formula_value_twins": len(audited_legs) == 3,
            "all_cells_reconstructed_from_embedded_sources": all(
                leg["workbook"]["comparison"][
                    "comparison_rows_reconstructed_from_sources"
                ] for leg in audited_legs.values()
            ),
            "all_returned_persisted_workbook_counts_exact": all(
                leg["publication"]["returned_persisted_workbook_counts_exact"]
                for leg in audited_legs.values()
            ),
            "all_duplicate_assignments_independently_exact": all(
                leg["publication"]["pairing_trace"][
                    "all_assignments_independently_exact_and_lexicographic"
                ] for leg in audited_legs.values()
            ),
            "source_semantic_pdf_excel_60493_0_1": (
                same_source["source_semantic_alignment"]
                == SOURCE_SEMANTIC_PDF_VS_EXCEL
            ),
            "product_suffix_key_anomaly_59946_547_548_reproduced": (
                same_source["product_key_alignment"] == {
                    "paired_rows": 59_946, "side_a_only_rows": 547,
                    "side_b_only_rows": 548,
                }
            ),
            "product_pdf_excel_source_contract_satisfied": False,
        },
    }
    if not all(
        value for key, value in result["invariants"].items()
        if key != "product_pdf_excel_source_contract_satisfied"
    ) or result["invariants"]["product_pdf_excel_source_contract_satisfied"] is not False:
        raise AuditError("final Highway Sequence witness invariants drift")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_bytes(_json_line(result))
    print(
        "PASS Highway Sequence product comparison twins (expected CMP-AUD-199 "
        f"defect reproduced): {args.output}"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AuditError as exc:
        print(f"FAIL Highway Sequence product comparison twin audit: {exc}")
        raise SystemExit(1)
