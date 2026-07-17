"""CMP-AUD-157 / CMP-AUD-045-HL: the TSN Highway Log v5 normalizer.

Hermetic, fixture-PDF-driven pins for the parser and consolidator:

  * a detached route-suffix group header ("07 LA 005 S") owns its rows as
    the suffixed TSMIS route (005S), never merged into the base route, and
    an unrecognized 4th token refuses;
  * asterisk-leading printed Descriptions inside the description band are
    conserved (the pre-v5 parser dropped "**** CODE ACCIDENTS TO" as a
    totals line), while real totals lines still close the open row;
  * the three per-row ADT claims are captured token-wise around the P/S
    flag (immune to the Look Back window overhang) and bound by digest;
  * every totals block is typed; TOTAL == CONST + UNCONST and all-zero
    suffixed-section totals are hard gates; a violation refuses publication;
  * report identity (OTM id, band date, title, cover year) is conserved and
    must agree across pages and members;
  * an unclassifiable below-band line refuses (zero unexplained residue);
  * a source-only mutation (one ADT digit) leaves the normalized rows
    byte-identical while the claims digest moves (the 157 mutation test);
  * consolidate() publishes the marker sheet on the per-route and combined
    workbooks and rides the claims on producer_extra.

The fixture PDFs are hand-rolled positioned-Helvetica documents parsed by
the production pdfplumber pipeline — no real corpus involved.
"""
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import logging
logging.getLogger("pdfminer").setLevel(logging.ERROR)

import consolidate_tsn_highway_log as HL  # noqa: E402
from events import Events  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

failures = []


def check(label, cond, detail=""):
    print(("OK   " if cond else "FAIL ") + label + (f" — {detail}" if detail and not cond else ""))
    if not cond:
        failures.append(label)


# --------------------------------------------------------------------------- #
# fixture-PDF writer (positioned Helvetica text; pdfplumber-parseable) —
# shared with check_compare_physical_identity via build/_hl_fixture_pdf.py
# --------------------------------------------------------------------------- #
sys.path.insert(0, str(ROOT / "build"))
from _hl_fixture_pdf import (band, cover, data_row, desc, group,  # noqa: E402
                             make_pdf, totals)


def parse(pdf_path, name):
    return HL.parse_pdf(pdf_path, Events(), pdf_name=name)


TMP = Path(tempfile.mkdtemp(prefix="tsmis_hl_claims_"))

# --------------------------------------------------------------------------- #
print("suffixed-route identity (CMP-AUD-045-HL): detached 4th token owns rows")
runs = (band(1)
        + group("07", "LA", "005", "S")
        + data_row(100, "001.000", "1.000", "000.000",
                   adt=[(452, "S"), (465, "24600")])
        + desc(112, "SUFFIX FEATURE")
        + totals(130, "county", total="000.000", const="000.000")
        + totals(150, "route", total="000.000", const="000.000")
        + group("07", "LA", "005", top=180)
        + data_row(200, "002.000", "0.500", "001.000",
                   adt=[(452, "S"), (465, "5000")])
        + desc(212, "BASE FEATURE")
        + totals(230, "county", total="000.500", const="000.500")
        + totals(250, "route", total="000.500", const="000.500"))
p = TMP / "D07 suffix.pdf"
make_pdf(p, [cover("07"), runs])
district, routes, claims = parse(p, "D07 suffix.pdf")
check("suffixed rows own route 005S; base rows own 005",
      sorted(routes) == ["005", "005S"]
      and [r.get("location") for r in routes["005S"]] == ["001.000"]
      and [r.get("location") for r in routes["005"]] == ["002.000"])
check("ownership manifest conserves district/county/route_token/suffix/route/rows",
      claims["ownership"] == [
          {"page": 2, "district": "07", "county": "LA", "route_token": "005",
           "suffix": "S", "route": "005S", "n_rows": 1},
          {"page": 2, "district": "07", "county": "LA", "route_token": "005",
           "suffix": "", "route": "005", "n_rows": 1}])
check("suffixed section's all-zero totals verify (hard gate bucket clean)",
      claims["totals"]["reconciliation"]["suffixed_zero"]["checked"] == 2
      and claims["totals"]["reconciliation"]["suffixed_zero"]["mismatches"] == [])
check("base route totals reconcile exactly once the suffix rows are separated",
      claims["totals"]["reconciliation"]["route"]["exact"] == 1
      and claims["totals"]["reconciliation"]["county"]["exact"] == 1)

bad = TMP / "D07 bad4th.pdf"
make_pdf(bad, [cover("07"),
               band(1) + group("07", "LA", "005", "SX")
               + data_row(100, "001.000", "1.000", "000.000")])
try:
    parse(bad, "D07 bad4th.pdf")
    check("a non-single-letter 4th group token refuses", False)
except ValueError as e:
    check("a non-single-letter 4th group token refuses",
          "group-header grammar" in str(e))

# --------------------------------------------------------------------------- #
print("asterisk-leading Descriptions conserved; totals still close the row")
runs = (band(1)
        + group("01", "MEN", "101")
        + data_row(100, "001.000", "1.000", "000.000")
        + desc(112, "**** CODE ACCIDENTS TO")
        + data_row(130, "002.000", "0.500", "001.000")
        + desc(142, "*")
        + totals(160, "volume")
        + desc(172, "AFTER TOTALS ORPHAN")          # residue: totals closed the row
        )
p2 = TMP / "D01 stars.pdf"
make_pdf(p2, [cover("01"), runs])
try:
    parse(p2, "D01 stars.pdf")
    check("a description-band line after a totals line (no open row) refuses "
          "as residue", False)
except ValueError as e:
    check("a description-band line after a totals line (no open row) refuses "
          "as residue", "could not be classified" in str(e))

runs_ok = (band(1)
           + group("01", "MEN", "101")
           + data_row(100, "001.000", "1.000", "000.000")
           + desc(112, "**** CODE ACCIDENTS TO")
           + data_row(130, "002.000", "0.500", "001.000")
           + desc(142, "*")
           + totals(160, "volume"))
p2b = TMP / "D01 stars ok.pdf"
make_pdf(p2b, [cover("01"), runs_ok])
_d, routes2, claims2 = parse(p2b, "D01 stars ok.pdf")
check("'**** CODE ACCIDENTS TO' and a bare '*' are conserved as Descriptions",
      [r.get("description") for r in routes2["101"]]
      == ["**** CODE ACCIDENTS TO", "*"])
check("the star totals line still parsed as a volume block",
      claims2["totals"]["kind_counts"].get("volume") == 1)

# --------------------------------------------------------------------------- #
print("ADT claims: token-wise around the P/S flag; digest-bound")
runs = (band(1)
        + group("06", "KER", "178")
        # wide Look Back overhangs the 448 window boundary (the pre-v5
        # window-center split would corrupt it into '24,00' / '0 P')
        + data_row(100, "001.000", "1.000", "000.000",
                   adt=[(426, "24,000"), (452, "P"), (465, "31500")])
        + data_row(130, "002.000", "0.500", "001.000",
                   adt=[(452, "S"), (465, "5000")])
        + data_row(160, "003.000", "0.250", "001.500",
                   adt=[(452, "P"), (465, "D-C")]))
p3 = TMP / "D06 adt.pdf"
make_pdf(p3, [cover("06"), runs])
_d, routes3, claims3 = parse(p3, "D06 adt.pdf")
adt = claims3["adt"]
check("Look Back overhanging the window boundary stays one verbatim token",
      adt["non_empty"] == {"back": 1, "flag": 3, "ahead": 3}
      and adt["flag_vocabulary"] == {"P": 2, "S": 1})
check("ADT stream is digest-bound", len(adt["digest_sha256"]) == 64)
check("non-numeric Look Ahead claims (D-C) are conserved verbatim, "
      "rows carry no ADT columns",
      all("adt_back" not in HL.ROW_KEYS for _ in [0])
      and adt["rows"] == 3)

p3b = TMP / "D06 adt mut.pdf"
runs_mut = [r if r[2] != "31500" else (r[0], r[1], "31501") for r in runs]
make_pdf(p3b, [cover("06"), runs_mut])
_d, routes3b, claims3b = parse(p3b, "D06 adt mut.pdf")


def written(routes):
    """The normalized output projection — exactly what the workbook rows
    carry (ROW_KEYS; the ADT zone never reaches a written column)."""
    return {r: [[row.get(k) for k in HL.ROW_KEYS] for row in rows]
            for r, rows in routes.items()}


check("a source-only ADT mutation leaves the normalized rows identical",
      written(routes3) == written(routes3b))
check("...but moves the claims digest (the CMP-AUD-157 mutation test)",
      claims3["adt"]["digest_sha256"] != claims3b["adt"]["digest_sha256"])

# --------------------------------------------------------------------------- #
print("hard gates: TCU arithmetic, residue, page/member identity")
runs = (band(1)
        + group("02", "SHA", "005")
        + data_row(100, "001.000", "1.000", "000.000")
        + totals(130, "county", total="002.000", const="001.000",
                 unconst="000.000"))              # 2 != 1 + 0
p4 = TMP / "D02 tcu.pdf"
make_pdf(p4, [cover("02"), runs])
_d, _r, claims4 = parse(p4, "D02 tcu.pdf")
check("a printed TOTAL != CONST + UNCONST lands in the gated tcu bucket",
      claims4["totals"]["reconciliation"]["tcu"]["mismatches"] != [])
check("_reconciliation_problems reports it (consolidate refuses publication)",
      "tcu" in HL._reconciliation_problems([claims4]))

p5 = TMP / "D02 residue.pdf"
make_pdf(p5, [cover("02"),
              band(1) + group("02", "SHA", "005")
              + data_row(100, "001.000", "1.000", "000.000")
              + [(300, 130, "UNEXPECTED MID-PAGE TEXT")]])
try:
    parse(p5, "D02 residue.pdf")
    check("an unclassifiable below-band line refuses (zero residue)", False)
except ValueError as e:
    check("an unclassifiable below-band line refuses (zero residue)",
          "could not be classified" in str(e))

p6 = TMP / "D02 dates.pdf"
make_pdf(p6, [cover("02"),
              band(1) + group("02", "SHA", "005")
              + data_row(100, "001.000", "1.000", "000.000"),
              band(2, date="09/16/25")
              + group("02", "SHA", "036", top=70)
              + data_row(100, "001.000", "1.000", "000.000")])
try:
    parse(p6, "D02 dates.pdf")
    check("pages disagreeing on the printed report date refuse", False)
except ValueError as e:
    check("pages disagreeing on the printed report date refuse",
          "report date" in str(e))

# --------------------------------------------------------------------------- #
print("page-break totals wraps: past-header continuation + stranded halves")
# The print reprints the current group header at each page top; a totals
# object wrapping across the break continues BELOW it (D02 p7->p8). The
# block must survive the header and reconcile against ITS OWN group.
page1 = (band(1)
         + group("02", "TRI", "003")
         + data_row(100, "001.000", "1.000", "000.000")
         + [(13.9, 130, "*** *** COUNTY TOTALS (MILEAGE) TOTAL 001.000 "
                        "CONST 001.000 UNCONST 000.000")])
page2 = (band(2)
         + group("02", "TRI", "003")                     # reprinted header
         + [(117.6, 100, "(DVMS) 1,234")]                # the block's tail
         + group("02", "SIS", "003", top=130)
         + data_row(160, "000.408", "2.329", "000.000")
         + [(13.9, 190, "*** *** COUNTY TOTALS (MILEAGE) TOTAL 002.329 "
                        "CONST 002.329 UNCONST 000.000")]
         + data_row(220, "002.737", "0.100", "002.329")  # closes the block
         + [(170.4, 250, "TOTAL CONST UNCONST")])        # stranded half
p7 = TMP / "D02 pagewrap.pdf"
make_pdf(p7, [cover("02"), page1, page2])
_d, _r, claims7 = parse(p7, "D02 pagewrap.pdf")
rec7 = claims7["totals"]["reconciliation"]
check("a totals block survives the reprinted page-top group header and "
      "reconciles against its OWN county",
      rec7["county"]["checked"] == 2 and rec7["county"]["exact"] == 2,
      str(rec7["county"]))
check("the continued block carried its (DVMS) fragment",
      claims7["totals"]["kind_counts"].get("county") == 2)
check("a stranded totals keyword half is conserved as a stray fragment, "
      "not residue",
      claims7["totals"]["stray_fragments"] == [
          {"page": 3, "x0": 170.4, "text": "TOTAL CONST UNCONST"}],
      str(claims7["totals"]["stray_fragments"]))

# The one CENTERED star line — "*** End of Report ***" (x0~209, outside the
# left star zone) — is its own typed marker, never absorbed into an open
# block and never a description.
p8 = TMP / "D02 endmarker.pdf"
make_pdf(p8, [cover("02"),
              band(1) + group("02", "SHA", "005")
              + data_row(100, "001.000", "1.000", "000.000")
              + totals(130, "route", total="001.000", const="001.000")
              + [(208.6, 160, "*** End of Report ***")]])
_d, _r, claims8 = parse(p8, "D02 endmarker.pdf")
check("a centered '*** End of Report ***' records as its own typed block",
      claims8["totals"]["kind_counts"].get("end_of_report") == 1
      and claims8["totals"]["kind_counts"].get("route") == 1,
      str(claims8["totals"]["kind_counts"]))

# --------------------------------------------------------------------------- #
print("consolidate(): 12-district build publishes markers + claims")
raw = TMP / "raw"
raw.mkdir()
for d in range(1, 13):
    dd = f"{d:02d}"
    runs = (band(1)
            + group(dd, "LA", "005")
            + data_row(100, "001.000", "1.000", "000.000",
                       adt=[(452, "S"), (465, "5000")])
            + desc(112, f"FEATURE D{dd}")
            + totals(130, "county", total="001.000", const="001.000")
            + totals(150, "route", total="001.000", const="001.000"))
    make_pdf(raw / f"D{dd} Highway Log TSN.pdf", [cover(dd), runs])
out = TMP / "consolidated.xlsx"
res = HL.consolidate(events=Events(), confirm_overwrite=lambda p: True,
                     input_dir=raw, out_path=out)
check("12-district fixture build succeeds", res.status == "ok",
      getattr(res, "message", ""))
check("producer_extra rides the cross-member claims record",
      (getattr(res, "producer_extra", None) or {}).get("tsn_source_claims", {})
      .get("report_id") == "OTM52010")
sc = res.producer_extra["tsn_source_claims"]
check("cross-member record aggregates 12 documents and the identity fields",
      len(sc["documents"]) == 12 and sc["report_date"] == "09/15/25"
      and sc["cover_year"] == "2025")
check("summary names the print identity",
      any("OTM52010" in ln for ln in res.summary_lines))
wb = load_workbook(out, read_only=True)
has_marker = HL.MARKER_SHEET in wb.sheetnames
version_row = None
if has_marker:
    for r in wb[HL.MARKER_SHEET].iter_rows(values_only=True):
        if r and str(r[0]).strip() == "Normalization version":
            version_row = r[1]
wb.close()
check("the combined workbook carries the v5 marker sheet",
      has_marker and version_row == HL.NORMALIZATION_VERSION)

# a broken-arithmetic member refuses the whole publication
make_pdf(raw / "D01 Highway Log TSN.pdf",
         [cover("01"),
          band(1) + group("01", "MEN", "101")
          + data_row(100, "001.000", "1.000", "000.000")
          + totals(130, "county", total="009.000", const="001.000",
                   unconst="000.000")])
res2 = HL.consolidate(events=Events(), confirm_overwrite=lambda p: True,
                      input_dir=raw, out_path=TMP / "consolidated2.xlsx")
check("a member whose printed totals break TOTAL=CONST+UNCONST refuses "
      "publication", res2.status == "error"
      and "reconcile" in (res2.message or ""))

# --------------------------------------------------------------------------- #
import shutil  # noqa: E402
shutil.rmtree(TMP, ignore_errors=True)

if failures:
    print(f"\nFAILED: {len(failures)}")
    sys.exit(1)
print("\nTSN Highway Log v5 claims/identity (CMP-AUD-157/045-HL): PASS")
