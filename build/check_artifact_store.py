"""CT-4 / CT-5 / CT-8 / CT-8b + fingerprint — the transactional artifact lifecycle
(scripts/artifact_store.py, P2).

Exercises the leaf module's three concerns with fault injection — NOT just happy paths:
  * atomic single-file write (CT-8): an interrupted/failed write never truncates a prior
    good artifact; no temp-name leak;
  * multi-file commit (CT-8b / R1-T02): values-canonical / formulas-best-effort; a 1st
    (values) or 2nd (formulas) save failure leaves the right state; no temp name surfaces
    in the result; overwrite-confirm honored against the FINAL path;
  * journaled store promotion + startup recovery (CT-4 / CT-5 / R1-T01): validate-before-
    commit, restore-from-backup on a death between renames, clean stale residue,
    idempotent;
  * input fingerprint (R1-R03): a DELETED file changes the identity (which a newest-mtime
    signal misses); missing/corrupt sidecar => stale.

Real openpyxl (a valid workbook is what `commit_workbook` validates); no browser/network.
Run from the repo root:
    build\\.venv\\Scripts\\python.exe build\\check_artifact_store.py
"""
import contextlib
import hashlib
import os
import subprocess
import sys
import uuid
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]   # scripts + repo root (updater -> version)

import artifact_store as a            # noqa: E402
from comparison_contract import ComparisonCounts, ComparisonOutcome  # noqa: E402
from events import ConsolidateResult  # noqa: E402
from openpyxl import Workbook         # noqa: E402

_failures = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _failures.append(name)


def _typed_compare_result(path, *, completion="complete", summary_lines=None):
    counts = ComparisonCounts(known=True, paired_rows=1)
    verdict = "match" if completion == "complete" else "diff"
    typed = ComparisonOutcome(
        status="ok", completion=completion, verdict=verdict, counts=counts,
        pairing_quality="exact")
    return ConsolidateResult(
        status="ok", output_path=str(path),
        summary_lines=list(summary_lines or ()), verdict=verdict,
        completion=completion, comparison_outcome=typed)


# Ownership predicate for recover_promotions (P2-B04). The MECHANICS tests seed their own
# store roots, so they accept all shape-valid journals; the ownership test passes a realistic one.
def _OWN_ALL(_store_root, _target):
    return True


@contextlib.contextmanager
def _patch(obj, attr, value):
    sentinel = object()
    old = getattr(obj, attr, sentinel)
    setattr(obj, attr, value)
    try:
        yield
    finally:
        if old is sentinel:
            delattr(obj, attr)
        else:
            setattr(obj, attr, value if False else old)


class _OsShim:
    """A stand-in for artifact_store's `os` whose `replace` raises for a target path
    substring; everything else delegates to the real os (scoped to artifact_store)."""
    def __init__(self, fail_substr):
        self._fail = fail_substr

    def replace(self, src, dst):
        if self._fail and self._fail in str(dst):
            raise OSError(13, "simulated locked destination")
        return os.replace(src, dst)

    def __getattr__(self, name):
        return getattr(os, name)


def _xlsx(path, sheet=None, value="x"):
    wb = Workbook()
    if sheet:
        wb.active.title = sheet
    if sheet == "Comparison":
        # CMP-AUD-115: a typed comparison artifact has to satisfy the commit
        # boundary's comparison-artifact schema (uniquely labelled Status/Diffs
        # + a valid status on every row). `value` still varies the bytes so
        # fixtures can tell two commits apart.
        wb.active.append(["Route", "Status", "Diffs"])
        wb.active.append([str(value), "Both", 0])
    else:
        wb.active["A1"] = value
    wb.save(str(path))


def _malformed_xlsx(path):
    """A ZIP that CONTAINS ``xl/workbook.xml`` (so a name-only check would pass) but whose
    workbook part is not valid XML — openpyxl cannot open it (the P2-B05 repro)."""
    import zipfile
    with zipfile.ZipFile(str(path), "w") as z:
        z.writestr("[Content_Types].xml", "<Types/>")
        z.writestr("xl/workbook.xml", b"not xml")


def _workbook_with_bad_sheet(path):
    """A valid XLSX (valid ZIP + valid workbook.xml) whose ``xl/worksheets/sheet1.xml`` is
    corrupt XML — only OPENING the workbook detects it (the P2-R04 repro)."""
    import zipfile
    good = Path(str(path) + ".good.xlsx")
    _xlsx(good, sheet="Comparison")
    with zipfile.ZipFile(good) as zin:
        data = {n: zin.read(n) for n in zin.namelist()}
    member = next(n for n in data if n.startswith("xl/worksheets/") and n.endswith(".xml"))
    data[member] = b"<worksheet><unclosed"
    with zipfile.ZipFile(str(path), "w") as zout:
        for n, d in data.items():
            zout.writestr(n, d)
    good.unlink()


@contextlib.contextmanager
def _fail_rename_when(should_fail):
    """Temporarily make ``Path.rename`` raise for (self, target) pairs matching
    `should_fail`, delegating otherwise — to drive the rollback/recovery failure paths."""
    orig = Path.rename

    def _r(self, target):
        if should_fail(self, Path(target)):
            raise OSError(13, "simulated rename failure")
        return orig(self, target)
    Path.rename = _r
    try:
        yield
    finally:
        Path.rename = orig


def _no_temp_leak(folder):
    return not any(".tmp-" in p.name for p in Path(folder).iterdir())


# --------------------------------------------------------------------------- #
def test_atomic_save(tmp):                                    # CT-8
    print("CT-8 atomic_save — an interrupted write never truncates the prior file:")
    d = tmp / "atomic"; d.mkdir()
    final = d / "out.xlsx"
    _xlsx(final)
    prior = final.read_bytes()

    # A failed os.replace (destination locked) must leave the prior file intact + raise.
    wb = Workbook(); wb.active["A1"] = "new"
    raised = False
    with _patch(a, "os", _OsShim("out.xlsx")):
        try:
            a.atomic_save(wb, final)
        except OSError:
            raised = True
    check("a failed replace raises", raised)
    check("...the prior workbook is byte-for-byte preserved", final.read_bytes() == prior)
    check("...no temp file leaked", _no_temp_leak(d))

    # A clean save replaces atomically.
    wb2 = Workbook(); wb2.active["A1"] = "fresh"
    a.atomic_save(wb2, final)
    check("a clean save replaces the file", final.exists() and final.read_bytes() != prior)
    check("...still no temp leak", _no_temp_leak(d))


def test_commit_single(tmp):                                  # CT-8 (single-file commit)
    print("CT-8 commit_workbook (single) — validate + rewrite + preserve-on-failure:")
    d = tmp / "single"; d.mkdir()
    final = d / "cmp.xlsx"

    def produce_ok(t):
        _xlsx(t)
        return ConsolidateResult(status="ok", output_path=str(t),
                                 summary_lines=[f"Values file: {t}"])
    res = a.commit_workbook(final, produce_ok)
    check("status ok", res.status == "ok")
    check("output_path rewritten to the FINAL name (no temp)",
          res.output_path == str(final) and ".tmp-" not in res.output_path)
    check("summary line rewritten (no temp name)",
          all(".tmp-" not in s for s in res.summary_lines) and str(final) in res.summary_lines[0])
    check("the workbook exists at the final path", final.exists())
    check("no temp leak", _no_temp_leak(d))

    # A producer that writes an INVALID file is rejected (not committed over a prior).
    _xlsx(final); prior = final.read_bytes()

    def produce_garbage(t):
        Path(t).write_text("not a workbook")
        return ConsolidateResult(status="ok", output_path=str(t))
    res = a.commit_workbook(final, produce_garbage)
    check("an invalid produced workbook -> error result", res.status == "error")
    check("...the prior file is preserved", final.read_bytes() == prior)
    check("...no temp leak", _no_temp_leak(d))

    # A producer that returns a non-ok result commits nothing.
    def produce_err(t):
        _xlsx(t)
        return ConsolidateResult(status="error", message="boom")
    res = a.commit_workbook(final, produce_err)
    check("a non-ok producer result is returned untouched", res.status == "error")
    check("...the prior file is still preserved", final.read_bytes() == prior)
    check("...no temp leak", _no_temp_leak(d))

    # Overwrite confirm is checked against the FINAL path (declined -> cancelled).
    def produce_should_not_run(t):
        produce_should_not_run.ran = True            # noqa
        _xlsx(t)
        return ConsolidateResult(status="ok", output_path=str(t))
    produce_should_not_run.ran = False
    res = a.commit_workbook(final, produce_should_not_run, confirm_overwrite=lambda _p: False)
    check("confirm declined on an existing final -> cancelled", res.status == "cancelled")
    check("...the producer never ran", produce_should_not_run.ran is False)
    check("...the prior file is preserved", final.read_bytes() == prior)


def test_commit_twin(tmp):                                    # CT-8b / R1-T02
    print("CT-8b commit_workbook (mode=both) — values-canonical, formulas best-effort:")
    d = tmp / "twin"; d.mkdir()
    final = d / "cmp.xlsx"                            # the picked name = formulas primary
    final_values = a._values_twin(final)

    def produce_both(t):
        _xlsx(t)                                      # formulas (primary)
        _xlsx(a._values_twin(t))                      # values twin
        return ConsolidateResult(status="ok", output_path=str(t),
                                 summary_lines=[f"Live-formulas file: {t}",
                                                f"Values file: {a._values_twin(t)}"])

    # Happy path: both committed; BOTH returned paths rewritten; no temp leak.
    res = a.commit_workbook(final, produce_both, twin=True)
    check("status ok", res.status == "ok")
    check("formulas committed to the picked name", final.exists())
    check("values committed to the (values) sibling", final_values.exists())
    check("output_path rewritten (formulas final, no temp)",
          res.output_path == str(final) and ".tmp-" not in res.output_path)
    check("both summary paths rewritten (no temp names)",
          all(".tmp-" not in s for s in res.summary_lines)
          and any(str(final_values) in s for s in res.summary_lines))
    check("no temp leak", _no_temp_leak(d))

    # 1st save fails: the VALUES (transactional) commit fails -> error, prior kept, no leak.
    _xlsx(final); _xlsx(final_values)
    pf, pv = final.read_bytes(), final_values.read_bytes()
    with _patch(a, "os", _OsShim("(values)")):
        res = a.commit_workbook(final, produce_both, twin=True)
    check("values-commit failure -> error result", res.status == "error")
    check("...the prior values workbook is preserved", final_values.read_bytes() == pv)
    check("...the prior formulas workbook is preserved (untouched)", final.read_bytes() == pf)
    check("...no temp leak (both temps cleaned)", _no_temp_leak(d))

    # 2nd save fails: values commits; the FORMULAS sibling is best-effort -> ok, no leak.
    final.unlink(); final_values.unlink()
    with _patch(a, "os", _OsShim("cmp.xlsx")):       # fail only the formulas (picked-name) replace
        res = a.commit_workbook(final, produce_both, twin=True)
    check("formulas-commit failure -> still ok (values is canonical)", res.status == "ok")
    check("...the values workbook IS committed", final_values.exists())
    check("...the formulas workbook is NOT present (best-effort skipped)", not final.exists())
    check("...no temp leak", _no_temp_leak(d))


def _member_matches(member, path, flavor, role, expected_bytes):
    path = Path(path)
    stat = path.stat()
    relative = member.get("relative_path")
    return (
        set(member) == {
            "flavor", "path", "relative_path", "canonical_path_at_write",
            "commit_role", "sha256", "size", "mtime_ns",
        }
        and member["flavor"] == flavor
        and member["path"] == str(path)
        and relative == path.name
        and relative not in (".", "..")
        and "/" not in relative and "\\" not in relative and ":" not in relative
        and member["canonical_path_at_write"] == a._resolved_identity(path)
        and member["commit_role"] == role
        and member["sha256"] == hashlib.sha256(expected_bytes).hexdigest()
        and member["size"] == len(expected_bytes) == stat.st_size
        and member["mtime_ns"] == stat.st_mtime_ns)


def _valid_uuid(value):
    try:
        return str(uuid.UUID(value)) == value
    except (TypeError, ValueError, AttributeError):
        return False


def test_comparison_artifact_generation(tmp):                 # CMP-AUD-075 / Phase 2C
    print("CMP-AUD-075 typed comparison generations list exact committed members:")
    d = tmp / "comparison-generation"; d.mkdir()
    generation_ids = []

    # Both lone modes are explicit: commit_workbook must never infer flavor from
    # workbook cells, names, or summary prose.
    for flavor, completion in (("values", "complete"), ("formulas", "partial")):
        final = d / f"single-{flavor}.xlsx"
        expected = [None]

        def produce_single(t, _flavor=flavor, _completion=completion):
            _xlsx(t, sheet="Comparison", value=f"{_flavor}-bytes")
            expected[0] = Path(t).read_bytes()
            return _typed_compare_result(t, completion=_completion)

        res = a.commit_workbook(
            final, produce_single, expect_sheet="Comparison",
            requested_mode=flavor)
        generation = res.artifact_generation
        generation_ids.append(generation.generation_id)
        check(f"lone {flavor}: exact canonical member + read-only hash",
              res.status == "ok" and final.read_bytes() == expected[0]
              and generation.requested_mode == flavor
              and generation.completion == completion
              and generation.publication_state == "committed"
              and len(generation.members) == 1
              and _member_matches(generation.members[0], final, flavor,
                                  "canonical", expected[0]))
        check(f"lone {flavor}: generation and succeeded attempt share fresh UUID",
              _valid_uuid(generation.generation_id)
              and res.attempt_state.state == "succeeded"
              and res.attempt_state.generation_id == generation.generation_id)
        check(f"lone {flavor}: digest map mirrors the sole member",
              generation.content_digests
              == {flavor: generation.members[0]["sha256"]})

    # Successful both mode has a fixed values-first order even though the picked
    # path is the formulas name and compare_core writes formulas first.
    final = d / "both.xlsx"
    final_values = a._values_twin(final)
    expected = {}

    def produce_both(t):
        _xlsx(t, sheet="Comparison", value="formulas-bytes")
        _xlsx(a._values_twin(t), sheet="Comparison", value="values-bytes")
        expected["formulas"] = Path(t).read_bytes()
        expected["values"] = a._values_twin(t).read_bytes()
        return _typed_compare_result(
            t, summary_lines=[f"Live-formulas file: {t}",
                              f"Values file: {a._values_twin(t)}"])

    res = a.commit_workbook(
        final, produce_both, twin=True, expect_sheet="Comparison",
        requested_mode="both")
    generation = res.artifact_generation
    generation_ids.append(generation.generation_id)
    check("both: canonical order is values then best-effort formulas",
          [member["flavor"] for member in generation.members]
          == ["values", "formulas"]
          and _member_matches(generation.members[0], final_values, "values",
                              "canonical", expected["values"])
          and _member_matches(generation.members[1], final, "formulas",
                              "best_effort", expected["formulas"]))
    check("both: hashing did not alter either committed workbook",
          final_values.read_bytes() == expected["values"]
          and final.read_bytes() == expected["formulas"])
    check("both: exact digest map + shared succeeded attempt UUID",
          generation.content_digests == {
              member["flavor"]: member["sha256"] for member in generation.members}
          and _valid_uuid(generation.generation_id)
          and res.attempt_state.generation_id == generation.generation_id)
    check("every committed generation gets a fresh UUID",
          len(generation_ids) == len(set(generation_ids)))

    # A legacy/untyped producer remains behavior-compatible and never receives a
    # synthetic generation merely because a requested_mode was supplied.
    legacy = d / "legacy.xlsx"

    def produce_legacy(t):
        _xlsx(t, sheet="Comparison")
        return ConsolidateResult(status="ok", output_path=str(t))

    legacy_res = a.commit_workbook(
        legacy, produce_legacy, expect_sheet="Comparison", requested_mode="values")
    check("untyped legacy result gets no generation/attempt metadata",
          legacy_res.status == "ok" and legacy_res.artifact_generation is None
          and legacy_res.attempt_state is None)


def test_comparison_generation_failure_paths(tmp):             # Phase 2C adversarial
    print("CMP-AUD-075 generation failure paths are exact and fail closed:")
    d = tmp / "comparison-generation-failures"; d.mkdir()

    def typed_both(t):
        _xlsx(t, sheet="Comparison", value="formulas")
        _xlsx(a._values_twin(t), sheet="Comparison", value="values")
        return _typed_compare_result(
            t, summary_lines=[f"Live-formulas file: {t}",
                              f"Values file: {a._values_twin(t)}"])

    # Best-effort formulas replace fails: the generation truthfully contains the
    # canonical values member only, even if a stale formulas destination existed.
    final = d / "formulas-fail.xlsx"
    final_values = a._values_twin(final)
    _xlsx(final, value="stale-formulas")
    stale_formulas = final.read_bytes()
    with _patch(a, "os", _OsShim("formulas-fail.xlsx")):
        res = a.commit_workbook(
            final, typed_both, twin=True, expect_sheet="Comparison",
            requested_mode="both")
    generation = res.artifact_generation
    values_bytes = final_values.read_bytes()
    check("formulas failure: values-only committed generation",
          res.status == "ok" and final.read_bytes() == stale_formulas
          and len(generation.members) == 1
          and _member_matches(generation.members[0], final_values, "values",
                              "canonical", values_bytes)
          and generation.content_digests == {
              "values": generation.members[0]["sha256"]})

    # Transactional values failure commits no member and therefore creates no
    # generation or succeeded attempt. Existing files remain byte exact.
    primary = d / "primary-fail.xlsx"
    primary_values = a._values_twin(primary)
    _xlsx(primary, value="old-formulas"); _xlsx(primary_values, value="old-values")
    prior_formula, prior_values = primary.read_bytes(), primary_values.read_bytes()
    with _patch(a, "os", _OsShim("(values)")):
        failed = a.commit_workbook(
            primary, typed_both, twin=True, expect_sheet="Comparison",
            requested_mode="both")
    check("primary failure: no generation and prior bytes preserved",
          failed.status == "error" and failed.artifact_generation is None
          and failed.attempt_state is None
          and primary.read_bytes() == prior_formula
          and primary_values.read_bytes() == prior_values)

    # Typed mode absence/mismatch is rejected before publication. Cover all
    # ambiguous pairings rather than silently defaulting lone output to formulas.
    invalid_cases = ((False, None), (False, "both"), (True, "values"))
    for index, (twin, requested_mode) in enumerate(invalid_cases):
        target = d / f"invalid-{index}.xlsx"
        _xlsx(target, value="prior")
        prior = target.read_bytes()

        def invalid_producer(t, _twin=twin):
            _xlsx(t, sheet="Comparison", value="new")
            if _twin:
                _xlsx(a._values_twin(t), sheet="Comparison", value="new-values")
            return _typed_compare_result(t)

        invalid = a.commit_workbook(
            target, invalid_producer, twin=twin, expect_sheet="Comparison",
            requested_mode=requested_mode)
        check(f"invalid typed mode {index}: error/no generation/no publication",
              invalid.status == "error" and invalid.artifact_generation is None
              and invalid.attempt_state is None and target.read_bytes() == prior)

    # Values commits and is safely hashed; the exact formulas target then loses
    # its guard. The terminal result stays error/values-canonical and may claim
    # only the already-bound values member.
    guarded = d / "post-values-guard.xlsx"
    guarded_values = a._values_twin(guarded)

    def guard(path, **_binding):
        path = Path(path)
        if path == guarded and guarded_values.exists():
            return False
        return path == d or path.parent == d

    guarded_result = a.commit_workbook(
        guarded, typed_both, twin=True, expect_sheet="Comparison",
        requested_mode="both", commit_guard=guard)
    generation = guarded_result.artifact_generation
    check("post-values guard loss: error points at values and invents no formulas",
          guarded_result.status == "error"
          and guarded_result.output_path == str(guarded_values)
          and guarded_values.exists() and not guarded.exists()
          and generation is not None and len(generation.members) == 1
          and generation.members[0]["flavor"] == "values"
          and generation.members[0]["relative_path"] == guarded_values.name)

    # An external byte mutation during hashing is detected by the fd/path
    # size+mtime binding. No digest/generation is published for mixed bytes.
    raced = d / "hash-race.xlsx"
    real_sha256 = a.hashlib.sha256
    did_mutate = [False]

    class RacingDigest:
        def __init__(self, *args, **kwargs):
            self.inner = real_sha256(*args, **kwargs)

        def update(self, data):
            if not did_mutate[0]:
                did_mutate[0] = True
                with raced.open("ab") as stream:
                    stream.write(b"external-race")
            self.inner.update(data)

        def hexdigest(self):
            return self.inner.hexdigest()

    with _patch(a.hashlib, "sha256", RacingDigest):
        raced_result = a.commit_workbook(
            raced,
            lambda t: (_xlsx(t, sheet="Comparison")
                       or _typed_compare_result(t)),
            expect_sheet="Comparison", requested_mode="values")
    check("hash mutation: committed file remains but no generation is certified",
          did_mutate[0] and raced_result.status == "error" and raced.exists()
          and raced_result.artifact_generation is None)

    # The file remains byte exact when only the post-hash ownership guard changes;
    # generation publication still fails closed.
    post_guard = d / "post-hash-guard.xlsx"
    expected = [None]
    hashed = [False]

    class MarkingDigest:
        def __init__(self, *args, **kwargs):
            self.inner = real_sha256(*args, **kwargs)

        def update(self, data):
            hashed[0] = True
            self.inner.update(data)

        def hexdigest(self):
            return self.inner.hexdigest()

    def hash_guard(path, **_binding):
        path = Path(path)
        return not (path == post_guard and hashed[0])

    def post_guard_producer(t):
        _xlsx(t, sheet="Comparison", value="guarded-bytes")
        expected[0] = Path(t).read_bytes()
        return _typed_compare_result(t)

    with _patch(a.hashlib, "sha256", MarkingDigest):
        post_guard_result = a.commit_workbook(
            post_guard, post_guard_producer, expect_sheet="Comparison",
            requested_mode="values", commit_guard=hash_guard)
    check("post-hash guard loss: bytes unchanged and no generation certified",
          hashed[0] and post_guard_result.status == "error"
          and post_guard.read_bytes() == expected[0]
          and post_guard_result.artifact_generation is None)


def test_commit_target_guard(tmp):                            # CMP-AUD-090 / S2
    print("S2 commit_workbook target guard — exact paths + bound temp identity:")
    d = tmp / "commit-target-guard"; d.mkdir()
    final = d / "comparison.xlsx"
    seen = []
    produced = [None]
    reserved_identity = [None]

    def guard(path, **binding):
        p = Path(path)
        seen.append((p, dict(binding)))
        return p == d or p.parent == d

    def produce(t):
        t = Path(t)
        produced[0] = t
        reserved_identity[0] = a._plain_entry_identity(t, directory=False)
        check("producer receives an already-reserved ordinary temp",
              reserved_identity[0] is not None and t.is_file())
        _xlsx(t, sheet="Comparison")
        check("serializer truncates the same reserved object (identity is stable)",
              a._plain_entry_identity(t, directory=False) == reserved_identity[0])
        return ConsolidateResult(status="ok", output_path=str(t))

    res = a.commit_workbook(
        final, produce, expect_sheet="Comparison", commit_guard=guard)
    guarded_paths = [p for p, _binding in seen]
    temp_bindings = [binding for p, binding in seen if p == produced[0]]
    check("guarded commit succeeds", res.status == "ok" and final.is_file())
    check("guard receives parent, exact final, and unpredictable exact temp",
          d in guarded_paths and final in guarded_paths and produced[0] in guarded_paths)
    check("temp checks carry the bound parent identity when the guard supports it",
          any(binding.get("anchor_path") == d
              and binding.get("anchor_identity") is not None
              for binding in temp_bindings))
    check("successful guarded commit leaves no temp", _no_temp_leak(d))

    # A target-aware guard may reject the unpredictable temp even though it
    # accepts the selected final.  Production must never start in that case.
    prior = final.read_bytes()
    ran = [False]

    def deny_temp(path, **_binding):
        return ".tmp-" not in Path(path).name

    def should_not_run(_t):
        ran[0] = True
        return ConsolidateResult(status="ok")

    denied = a.commit_workbook(
        final, should_not_run, expect_sheet="Comparison",
        commit_guard=deny_temp)
    check("a denied unpredictable temp fails before producer entry",
          denied.status == "error" and not ran[0])
    check("...prior final is byte-exact and no unowned temp was touched",
          final.read_bytes() == prior and _no_temp_leak(d))

    zero_arg_ran = [False]

    def zero_arg_producer(_t):
        zero_arg_ran[0] = True
        return ConsolidateResult(status="ok")

    zero_arg = a.commit_workbook(
        final, zero_arg_producer, commit_guard=lambda: True)
    check("a zero-argument guard cannot authorize an exact descendant target",
          zero_arg.status == "error" and not zero_arg_ran[0]
          and final.read_bytes() == prior)

    # Replace the reserved temp with another ordinary workbook.  A mere
    # safe-descendant check still passes, so the transaction's own file-ID
    # binding must reject it and must NOT unlink the unrelated replacement.
    replaced_temp = [None]

    def replace_reserved(t):
        t = Path(t)
        replacement = d / "replacement.xlsx"
        _xlsx(replacement, sheet="Comparison")
        os.replace(replacement, t)
        replaced_temp[0] = t
        return ConsolidateResult(status="ok", output_path=str(t))

    replaced = a.commit_workbook(
        final, replace_reserved, expect_sheet="Comparison", commit_guard=guard)
    check("a same-path ordinary temp replacement is never published",
          replaced.status == "error" and final.read_bytes() == prior)
    check("...the unrelated replacement is retained, never cleanup-unlinked",
          replaced_temp[0] is not None and replaced_temp[0].is_file())
    replaced_temp[0].unlink()

    # Loss after validation must be observed at the immediate publication
    # boundary.  Because authority is gone, the still-bound temp is retained
    # instead of being deleted through a potentially changed pathname.
    allow = [True]
    late_temp = [None]

    def late_guard(_path, **_binding):
        return allow[0]

    def valid_then_revoke(p):
        late_temp[0] = Path(p)
        ok = a._openable_xlsx(p, "Comparison")
        allow[0] = False
        return ok

    late = a.commit_workbook(
        final, produce, expect_sheet="Comparison", validate=valid_then_revoke,
        commit_guard=late_guard)
    check("guard loss during validation blocks the final os.replace",
          late.status == "error" and final.read_bytes() == prior)
    check("...post-loss cleanup retains the temp rather than mutating blindly",
          late_temp[0] is not None and late_temp[0].is_file())
    allow[0] = True
    late_temp[0].unlink()


def test_output_source_alias_guard(tmp):                       # CMP-AUD-041 / S1
    print("S1 output/source alias guard — direct, resolved, hardlink, and derived twins:")
    d = tmp / "source-alias"; d.mkdir()

    def producer(t):
        producer.calls += 1
        _xlsx(t, sheet="Comparison")
        return ConsolidateResult(status="ok", output_path=str(t))

    def commit(final, sources, *, twin=False, produce=producer, validate=None):
        """Compatibility wrapper: before S1 lands the unknown keyword is the RED."""
        try:
            kwargs = {"twin": twin, "expect_sheet": "Comparison",
                      "source_paths": sources}
            if validate is not None:
                kwargs["validate"] = validate
            return a.commit_workbook(final, produce, **kwargs)
        except TypeError:
            return None

    # The picked destination itself is a selected comparison input.  Overwrite
    # confirmation is irrelevant: a comparison may never destroy its source.
    direct = d / "selected-source.xlsx"
    _xlsx(direct, sheet="SourceData")
    direct_prior = direct.read_bytes()
    producer.calls = 0
    res = commit(direct, [direct])
    check("a direct output/source alias fails closed before production",
          res is not None and res.status == "error" and producer.calls == 0)
    check("...the selected source is byte-for-byte preserved",
          direct.read_bytes() == direct_prior)

    # Canonicalization must defeat relative/dot aliases even when the path text
    # differs.  (Path.resolve also covers symlinks/junctions where available.)
    # Build the alternate spelling on `direct`'s OWN drive (a `..` round-trip
    # through its parent) rather than relative to cwd: on CI the checkout and the
    # temp dir sit on different drives (D:\a\... vs C:\...\Temp), and a
    # cross-mount os.path.relpath raises "path is on mount 'C:', start on 'D:'".
    producer.calls = 0
    spelled_differently = direct.parent / ".." / direct.parent.name / direct.name
    res = commit(direct.resolve(), [spelled_differently])
    check("resolved aliases are rejected independent of path spelling",
          res is not None and res.status == "error" and producer.calls == 0)
    producer.calls = 0
    case_spelling = direct.with_name(direct.name.upper())
    res = commit(direct, [case_spelling])
    check("case-folded aliases are rejected on the Windows artifact contract",
          res is not None and res.status == "error" and producer.calls == 0)

    link_dest = d / "selected-source-link.xlsx"
    try:
        link_dest.symlink_to(direct)
    except OSError:
        check("symlink resolution probe skipped only when link creation is refused", True)
    else:
        producer.calls = 0
        res = commit(link_dest, [direct])
        check("a resolved symlink destination cannot overwrite its source target",
              res is not None and res.status == "error" and producer.calls == 0)

    # mode=both derives an unselected '(values)' destination.  A safe picked
    # formulas name must not let that sibling overwrite either input.
    formulas = d / "comparison.xlsx"
    values_source = a._values_twin(formulas)
    _xlsx(values_source, sheet="SourceData")
    values_prior = values_source.read_bytes()
    producer.calls = 0
    res = commit(formulas, [values_source], twin=True)
    check("a derived values twin that aliases a source is rejected",
          res is not None and res.status == "error" and producer.calls == 0)
    check("...the unselected sibling source is preserved",
          values_source.read_bytes() == values_prior and not formulas.exists())

    folder_source = d / "selected-export-folder"
    folder_source.mkdir()
    _xlsx(folder_source / "route.xlsx", sheet="SourceData")
    producer.calls = 0
    nested_output = folder_source / "comparison.xlsx"
    res = commit(nested_output, [folder_source])
    check("an output inside a selected source folder is rejected",
          res is not None and res.status == "error" and producer.calls == 0
          and not nested_output.exists())

    # Existing hardlinks have different canonical path text but the same file
    # identity.  Windows and CI filesystems normally support these; retain an
    # explicit feasibility branch for unusual filesystems.
    hard_source = d / "hard-source.xlsx"
    hard_dest = d / "hard-destination.xlsx"
    _xlsx(hard_source, sheet="SourceData")
    hard_prior = hard_source.read_bytes()
    try:
        os.link(hard_source, hard_dest)
    except OSError:
        check("hardlink identity probe skipped only when the filesystem refuses links", True)
    else:
        producer.calls = 0
        res = commit(hard_dest, [hard_source])
        check("os.path.samefile identity rejects an existing hardlink alias",
              res is not None and res.status == "error" and producer.calls == 0)
        check("...the hardlinked source content is preserved",
              hard_source.read_bytes() == hard_prior)

    def identity_denied(_left, _right):
        raise PermissionError(13, "identity unavailable")

    producer.calls = 0
    with _patch(a.os.path, "samefile", identity_denied):
        res = commit(d / "identity-unverified.xlsx", [hard_source])
    check("an unreadable same-file identity fails closed",
          res is not None and res.status == "error" and producer.calls == 0)

    # Re-check immediately before commit: an initially absent output can appear
    # as a hardlink to an input while the producer is running.
    late_source = d / "late-source.xlsx"
    late_dest = d / "late-destination.xlsx"
    _xlsx(late_source, sheet="SourceData")
    late_prior = late_source.read_bytes()

    def produce_late_alias(t):
        _xlsx(t, sheet="Comparison")
        os.link(late_source, late_dest)
        return ConsolidateResult(status="ok", output_path=str(t))

    try:
        res = commit(late_dest, [late_source], produce=produce_late_alias)
    except OSError:
        check("late hardlink race probe skipped only when links are unavailable", True)
    else:
        check("a destination that becomes a source alias during production is rejected",
              res is not None and res.status == "error")
        check("...the late-linked source is preserved and no temp leaks",
              late_source.read_bytes() == late_prior and _no_temp_leak(d))

    # The commit boundary performs one final identity check after workbook
    # validation, immediately before os.replace.  This closes a still-later race
    # where the destination changes while the produced workbook is being opened.
    commit_source = d / "commit-race-source.xlsx"
    commit_dest = d / "commit-race-destination.xlsx"
    _xlsx(commit_source, sheet="SourceData")
    commit_prior = commit_source.read_bytes()

    def validate_then_alias(_temp):
        os.link(commit_source, commit_dest)
        return True

    try:
        res = commit(commit_dest, [commit_source], validate=validate_then_alias)
    except OSError:
        check("pre-replace hardlink race probe skipped only when links are unavailable", True)
    else:
        check("an alias introduced during validation is rejected before os.replace",
              res is not None and res.status == "error")
        check("...the commit-race source is preserved and no temp leaks",
              commit_source.read_bytes() == commit_prior and _no_temp_leak(d))

    # Capture the source object, not merely its pathname.  A source can be moved
    # onto an initially absent destination while the producer runs and a decoy
    # planted at the original path; a pathname-only post-check compares the
    # destination with the decoy and destroys the actual selected object.
    moved_source = d / "moved-source.xlsx"
    moved_dest = d / "moved-destination.xlsx"
    _xlsx(moved_source, sheet="OriginalSource")
    moved_prior = moved_source.read_bytes()

    def produce_source_swap(t):
        _xlsx(t, sheet="Comparison")
        os.replace(moved_source, moved_dest)
        _xlsx(moved_source, sheet="DecoyReplacement")
        return ConsolidateResult(status="ok", output_path=str(t))

    res = commit(moved_dest, [moved_source], produce=produce_source_swap)
    check("a moved source plus decoy is rejected by its captured file identity",
          res is not None and res.status == "error")
    check("...the originally selected object survives at its moved path",
          moved_dest.read_bytes() == moved_prior and _no_temp_leak(d))

    # Cleanup must be identity-aware too.  If the selected source is moved onto
    # the producer temp, rejecting the run must RETAIN that path rather than
    # treating it as disposable comparison residue.
    temp_source = d / "temp-swap-source.xlsx"
    temp_dest = d / "temp-swap-destination.xlsx"
    _xlsx(temp_source, sheet="OriginalTempSource")
    temp_prior = temp_source.read_bytes()
    moved_temp = [None]

    def produce_source_at_temp(t):
        moved_temp[0] = Path(t)
        os.replace(temp_source, t)
        _xlsx(temp_source, sheet="DecoyReplacement")
        return ConsolidateResult(status="ok", output_path=str(t))

    res = commit(temp_dest, [temp_source], produce=produce_source_at_temp)
    check("a captured source moved onto the producer temp is rejected",
          res is not None and res.status == "error")
    check("...unsafe temp cleanup retains the originally selected object",
          moved_temp[0].read_bytes() == temp_prior and not temp_dest.exists())
    moved_temp[0].unlink()

    # Directory sources need the same stable-object rule when the final file is
    # below a directory's NEW name.  The producer simulates a rename of the
    # selected source directory followed by a same-name decoy directory.
    source_dir = d / "source-directory"
    source_dir.mkdir()
    route = source_dir / "route.xlsx"
    _xlsx(route, sheet="OriginalRoute")
    route_prior = route.read_bytes()
    renamed_dir = d / "renamed-source-directory"
    nested_dest = renamed_dir / "comparison.xlsx"
    hold = d / "held-comparison.xlsx"

    def produce_directory_swap(t):
        _xlsx(t, sheet="Comparison")
        os.replace(t, hold)
        Path(t).parent.rmdir()
        os.replace(source_dir, renamed_dir)
        source_dir.mkdir()
        os.replace(hold, t)
        return ConsolidateResult(status="ok", output_path=str(t))

    res = commit(nested_dest, [source_dir], produce=produce_directory_swap)
    check("a destination below a renamed captured source directory is rejected",
          res is not None and res.status == "error")
    check("...the original directory contents survive and no temp leaks",
          (renamed_dir / "route.xlsx").read_bytes() == route_prior
          and not nested_dest.exists() and not hold.exists() and _no_temp_leak(d))


def test_fingerprint(tmp):                                    # CT-6 (unit) / R1-R03
    print("fingerprint — identity over (name,size,mtime); a DELETE changes it:")
    store = tmp / "store"; store.mkdir()
    (store / "r1.xlsx").write_bytes(b"a" * 10)
    (store / "r2.xlsx").write_bytes(b"b" * 20)
    fp0 = a.fingerprint(store)
    check("stable across calls", a.fingerprint(store) == fp0)

    (store / "r3.xlsx").write_bytes(b"c" * 5)
    fp_add = a.fingerprint(store)
    check("adding a file changes the fingerprint", fp_add != fp0)

    (store / "r3.xlsx").unlink()
    check("removing it returns to the original", a.fingerprint(store) == fp0)

    (store / "r2.xlsx").write_bytes(b"b" * 21)        # resize
    check("resizing a file changes the fingerprint", a.fingerprint(store) != fp0)

    fp_metadata = a.fingerprint(store)
    payload_name = (
        ".cmpv3-" + "a" * 64 + "-000000-" + "b" * 64
        + ".comparison-payload.zlib")
    (store / "~$r1.xlsx").write_bytes(b"x")
    (store / "r1.xlsx.outcome.json").write_text("{}")
    (store / payload_name).write_bytes(b"strict comparison metadata")
    fallback_payload_name = (
        ".cmpv3-" + "a" * 64 + "-000000-" + "b" * 64
        + "-f-" + "c" * 64 + "-" + "d" * 16
        + ".comparison-payload.zlib")
    (store / fallback_payload_name).write_bytes(b"strict fallback comparison metadata")
    slot_payload_name = (
        ".cmpv3-" + "a" * 64 + "-000000-" + "b" * 64
        + "-f-07.comparison-payload.zlib")
    (store / slot_payload_name).write_bytes(b"strict bounded-slot comparison metadata")
    # CMP-AUD-242: current builds write the SHORT 16-hex shape — it must be
    # excluded from fingerprints exactly like the legacy full-hex names above.
    short_payload_name = (
        ".cmpv3-" + "a" * 16 + "-000000-" + "b" * 16
        + ".comparison-payload.zlib")
    (store / short_payload_name).write_bytes(b"strict short comparison metadata")
    short_slot_name = (
        ".cmpv3-" + "a" * 16 + "-000000-" + "b" * 16
        + "-f-00.comparison-payload.zlib")
    (store / short_slot_name).write_bytes(b"strict short-slot comparison metadata")
    publication_lock = store / a._COMPARISON_PUBLICATION_LOCK_NAME
    publication_lock.write_bytes(b"")
    check("excludes lock/outcome/fingerprint and exact primary/legacy/slot v3 metadata",
          a.fingerprint(store) == fp_metadata)
    check("permanent publication lock prevents parent disposal despite fingerprint exclusion",
          a._dir_is_disposable_placeholder(store) is False)
    lock_only = tmp / "publication-lock-only"
    lock_only.mkdir()
    (lock_only / a._COMPARISON_PUBLICATION_LOCK_NAME).write_bytes(b"")
    check("even a lock-only directory is never classified as disposable",
          a._dir_is_disposable_placeholder(lock_only) is False)
    near_payload = store / (
        ".cmpv3-" + "a" * 64 + "-000000-" + "b" * 63 + "G"
        + ".comparison-payload.zlib")
    near_payload.write_bytes(b"foreign near-match")
    check("near-match .zlib user file is not broadly excluded",
          a.fingerprint(store) != fp_metadata)
    near_payload.unlink()
    near_fallback = store / (
        ".cmpv3-" + "a" * 64 + "-000000-" + "b" * 64
        + "-f-" + "c" * 64 + "-" + "d" * 15 + "G"
        + ".comparison-payload.zlib")
    near_fallback.write_bytes(b"foreign fallback near-match")
    check("near-match fallback .zlib is not broadly excluded",
          a.fingerprint(store) != fp_metadata)
    near_fallback.unlink()
    near_slot = store / (
        ".cmpv3-" + "a" * 64 + "-000000-" + "b" * 64
        + "-f-08.comparison-payload.zlib")
    near_slot.write_bytes(b"foreign bounded-slot near-match")
    check("out-of-range fallback slot is not broadly excluded",
          a.fingerprint(store) != fp_metadata)
    check("an absent folder -> the UNREADABLE sentinel (caller rebuilds)",
          a.fingerprint(tmp / "nope") == a._UNREADABLE)


def test_consolidated_fresh(tmp):                             # CT-6 (the freshness gate)
    print("consolidated_fresh — a deleted route reads STALE (newest-mtime missed it):")
    store = tmp / "cstore"; store.mkdir()
    (store / "r1.xlsx").write_bytes(b"a" * 10)
    (store / "r2.xlsx").write_bytes(b"b" * 20)
    (store / "r3.xlsx").write_bytes(b"c" * 30)
    consolidated = tmp / "combined.xlsx"
    _xlsx(consolidated)

    check("no sidecar yet -> stale (legacy/one-time migration)",
          a.consolidated_fresh(consolidated, store) is False)
    a.write_consolidated_fingerprint(consolidated, store)
    check("after recording the fingerprint -> fresh", a.consolidated_fresh(consolidated, store))

    # Delete a NON-newest route: newest-mtime is unchanged, but identity changed.
    (store / "r1.xlsx").unlink()
    check("a DELETED route -> stale (the F5 fix)",
          a.consolidated_fresh(consolidated, store) is False)

    # Re-record, then a missing workbook is stale regardless of the sidecar.
    a.write_consolidated_fingerprint(consolidated, store)
    check("re-recorded -> fresh again", a.consolidated_fresh(consolidated, store))
    consolidated.unlink()
    check("a missing consolidated workbook -> stale", a.consolidated_fresh(consolidated, store) is False)
    _xlsx(consolidated)

    # A corrupt sidecar reads stale (never a false-fresh).
    a.write_consolidated_fingerprint(consolidated, store)
    a._fp_sidecar(consolidated).write_text("{ not json")
    check("a corrupt fingerprint sidecar -> stale", a.consolidated_fresh(consolidated, store) is False)


def _seed_store(parent, name="live", files=("r1.xlsx", "r2.xlsx")):
    d = parent / name
    d.mkdir(parents=True)
    for f in files:
        (d / f).write_bytes(b"data-" + f.encode())
    return d


def test_fingerprint_commit_guard(tmp):                       # CMP-AUD-090
    print("CMP-AUD-090 fingerprint publication guards exact temp/final paths:")
    store = tmp / "guarded-fingerprint-store"; store.mkdir()
    (store / "r1.xlsx").write_bytes(b"route")
    consolidated = tmp / "guarded_combined.xlsx"; _xlsx(consolidated)
    sidecar = a._fp_sidecar(consolidated)
    seen = []

    def _allow(path):
        seen.append(Path(path))
        return True

    ok = a.write_consolidated_fingerprint(
        consolidated, store, commit_guard=_allow)
    check("an allowed guarded sidecar publishes successfully", ok is True)
    check("the guard receives the exact final sidecar path", sidecar in seen)
    check("the guard also receives the unpredictable exact temp path",
          any(p.parent == sidecar.parent and p.name.startswith(sidecar.name + ".tmp-")
              for p in seen))

    prior = sidecar.read_bytes()
    seen.clear()

    def _deny_final(path):
        path = Path(path)
        seen.append(path)
        return path != sidecar

    ok = a.write_consolidated_fingerprint(
        consolidated, store, commit_guard=_deny_final)
    check("a denied final target fails closed before os.replace", ok is False)
    check("...the prior sidecar stays byte-exact and temp cleanup is guarded",
          sidecar.read_bytes() == prior
          and not list(sidecar.parent.glob(sidecar.name + ".tmp-*")))

    # The input-race invalidation ladder is mutating too. If ownership was lost,
    # it must not unlink/truncate a stale sidecar or quarantine a workbook at the
    # inherited pathname.
    before_workbook = consolidated.read_bytes()
    ok = a.write_consolidated_fingerprint(
        consolidated, store, built_from="different-input-identity",
        commit_guard=lambda _path: False)
    check("a denied race-invalidation target performs no fallback mutation",
          ok is False and sidecar.read_bytes() == prior
          and consolidated.read_bytes() == before_workbook)


def test_promote_validate(tmp):                               # CT-5
    print("CT-5 promote_store — validate before commit; locked live keeps last-good:")
    base = tmp / "p_validate"; base.mkdir()
    live = _seed_store(base, "live")
    prior = sorted(p.name for p in live.iterdir())

    # Empty staging is never promoted.
    empty_stage = base / "live.staging"; empty_stage.mkdir()
    check("empty staging -> not promoted", a.promote_store(live, empty_stage) is False)
    check("...live preserved", sorted(p.name for p in live.iterdir()) == prior)
    check("...empty staging cleaned", not empty_stage.exists())

    # Missing staging is never promoted.
    check("missing staging -> not promoted", a.promote_store(live, base / "gone.staging") is False)
    check("...live still preserved", sorted(p.name for p in live.iterdir()) == prior)

    # A locked live (the live->backup rename fails) keeps last-good, discards staging.
    stage = base / "live.staging2"; stage.mkdir()
    (stage / "new.xlsx").write_bytes(b"new")
    with _patch(a, "os", _OsShim("")):               # os passthrough; force the rename to fail
        orig_rename = Path.rename

        def _locked_rename(self, target):
            if self.name == "live" and ".bak-" in Path(target).name:
                raise OSError(13, "live locked")
            return orig_rename(self, target)
        with _patch(Path, "rename", _locked_rename):
            ok = a.promote_store(live, stage)
    check("a locked live -> not promoted (kept)", ok is False)
    check("...live still has the ORIGINAL files", sorted(p.name for p in live.iterdir()) == prior)
    check("...staging discarded", not stage.exists())

    # Clean promotion when nothing blocks it.
    stage3 = base / "live.staging3"; stage3.mkdir()
    (stage3 / "fresh.xlsx").write_bytes(b"fresh")
    check("a clean promotion succeeds", a.promote_store(live, stage3) is True)
    check("...live now holds the staged content", (live / "fresh.xlsx").exists()
          and not (live / "r1.xlsx").exists())
    # The journal FILE and backup are gone; the (possibly shared) empty `.promote` dir is
    # left for recover_promotions to rmdir (sibling stores share it — rmdir'ing it
    # mid-promotion would race a concurrent journal write).
    check("...no backup / journal residue left",
          not any(".bak-" in p.name for p in base.iterdir())
          and not list((base / ".promote").glob("*.json")))


def test_recovery(tmp):                                       # CT-4 / R1-T01
    print("CT-4 recover_promotions — seed each dead-process state; recovery repairs it:")
    import json

    # State A: death BETWEEN the renames — live missing, backup present, journal present.
    base = tmp / "rec_a"; base.mkdir()
    backup = _seed_store(base, "live.bak-tok")
    stage = base / "live.staging"; stage.mkdir(); (stage / "x").write_bytes(b"x")
    jdir = base / ".promote"; jdir.mkdir()
    (jdir / "tok.json").write_text(json.dumps(
        {"target": "live", "backup": "live.bak-tok", "staging": "live.staging", "token": "tok"}))
    a.recover_promotions(base, _OWN_ALL)
    check("live RESTORED from the backup", (base / "live").exists()
          and (base / "live" / "r1.xlsx").exists())
    check("...backup consumed", not backup.exists())
    check("...staging + journal cleaned", not stage.exists() and not (jdir / "tok.json").exists())
    check("...idempotent (a second sweep is a no-op)",
          (a.recover_promotions(base, _OWN_ALL), (base / "live").exists())[-1])

    # State B: promotion COMPLETED (live present) but residue remains — clean it, keep live.
    base = tmp / "rec_b"; base.mkdir()
    live = _seed_store(base, "live")
    stale_backup = _seed_store(base, "live.bak-old")
    jdir = base / ".promote"; jdir.mkdir()
    (jdir / "old.json").write_text(json.dumps(
        {"target": "live", "backup": "live.bak-old", "staging": "live.staging", "token": "old"}))
    a.recover_promotions(base, _OWN_ALL)
    check("live (present) is kept", (base / "live" / "r1.xlsx").exists())
    check("...stale backup cleaned", not stale_backup.exists())
    check("...journal removed", not (jdir / "old.json").exists())

    # State C (P2-B04): orphan *.staging / *.bak-* with NO journal are now LEFT UNTOUCHED —
    # `root` may be a user destination with unrelated folders, so the journal-free sweep that
    # guessed at app ownership was removed (harmless residue is preferable to deleting data).
    base = tmp / "rec_c"; base.mkdir()
    _seed_store(base, "live")
    orphan_stage = base / "live.staging"; orphan_stage.mkdir(); (orphan_stage / "x").write_bytes(b"x")
    orphan_bak = _seed_store(base, "live.bak-orphan")
    a.recover_promotions(base, _OWN_ALL)
    check("orphan staging with no journal is NOT swept (B04-safe)", orphan_stage.exists())
    check("orphan backup with no journal is NOT swept (B04-safe)", orphan_bak.exists())
    check("live untouched", (base / "live" / "r1.xlsx").exists())

    # State D: a corrupt journal is dropped, never raises.
    base = tmp / "rec_d"; base.mkdir()
    _seed_store(base, "live")
    jdir = base / ".promote"; jdir.mkdir()
    (jdir / "bad.json").write_text("{ not json")
    a.recover_promotions(base, _OWN_ALL)
    check("a corrupt journal is dropped (no raise)", not (jdir / "bad.json").exists())


def test_b05_malformed(tmp):                                  # P2-B05
    print("P2-B05 commit validation OPENS the workbook (rejects a malformed/unreadable part):")
    d = tmp / "b05"; d.mkdir()
    final = d / "cmp.xlsx"
    _xlsx(final, sheet="Comparison"); prior = final.read_bytes()

    # A ZIP that contains xl/workbook.xml but whose part is garbage — name-check would pass.
    def produce_malformed(t):
        _malformed_xlsx(t)
        return ConsolidateResult(status="ok", output_path=str(t))
    res = a.commit_workbook(final, produce_malformed, expect_sheet="Comparison")
    check("a ZIP-shaped but unreadable workbook -> error (not committed)", res.status == "error")
    check("...the valid prior workbook is preserved", final.read_bytes() == prior)
    check("...and it is still openable by openpyxl", a._openable_xlsx(final, "Comparison"))
    check("no temp leak", _no_temp_leak(d))

    # A readable workbook MISSING the expected sheet is rejected too.
    def produce_wrong_sheet(t):
        _xlsx(t, sheet="NotComparison")
        return ConsolidateResult(status="ok", output_path=str(t))
    res = a.commit_workbook(final, produce_wrong_sheet, expect_sheet="Comparison")
    check("a workbook missing the expected sheet -> error", res.status == "error")
    check("...prior preserved", final.read_bytes() == prior)


def test_r02_no_temp_leak(tmp):                               # P2-R02
    print("P2-R02 an ERROR result never leaks the deleted temp path (message/output/summary):")
    d = tmp / "r02"; d.mkdir()
    final = d / "cmp.xlsx"

    def produce_err_with_path(t):
        # compare_core builds its save-error message from the path it was handed (a temp).
        return ConsolidateResult(status="error", output_path=str(t),
                                 message=f"Could not save {Path(t).name}.",
                                 summary_lines=[f"Values file: {t}"])
    res = a.commit_workbook(final, produce_err_with_path)
    leaked = (".tmp-" in (res.message or "") or ".tmp-" in (res.output_path or "")
              or any(".tmp-" in s for s in (res.summary_lines or [])))
    check("the error result carries NO .tmp-<token> in any field", not leaked)
    check("...message names the FINAL file instead", final.name in (res.message or ""))
    check("no temp leak on disk", _no_temp_leak(d))


def test_r03_formulas_truthful(tmp):                          # P2-R03
    print("P2-R03 a best-effort formulas failure returns a TRUTHFUL values-canonical result:")
    d = tmp / "r03"; d.mkdir()
    final = d / "cmp.xlsx"                            # picked name = formulas primary
    final_values = a._values_twin(final)

    def produce_both(t):
        _xlsx(t, sheet="Comparison")                 # formulas
        _xlsx(a._values_twin(t), sheet="Comparison") # values twin
        return ConsolidateResult(status="ok", output_path=str(t),
                                 summary_lines=[f"Live-formulas file: {t}",
                                                f"Values file: {a._values_twin(t)}"])
    with _patch(a, "os", _OsShim("cmp.xlsx")):       # fail ONLY the formulas (picked-name) replace
        res = a.commit_workbook(final, produce_both, twin=True, expect_sheet="Comparison")
    check("status ok (values is canonical)", res.status == "ok")
    check("the values workbook IS committed", final_values.exists())
    check("the formulas workbook is NOT present", not final.exists())
    check("output_path points at the committed VALUES file (truthful)",
          res.output_path == str(final_values))
    check("the formulas summary line is a NOT-refreshed warning (no false claim)",
          any("NOT refreshed" in s for s in res.summary_lines)
          and not any(s.startswith(f"Live-formulas file: {final}") for s in res.summary_lines))
    check("no temp leak", _no_temp_leak(d))


def test_r01_directory_staging(tmp):                          # P2-R01
    print("P2-R01 a staging dir with no eligible report FILE is never promoted:")
    base = tmp / "r01"; base.mkdir()
    live = _seed_store(base, "live")
    prior = sorted(p.name for p in live.iterdir())
    stage = base / "live.staging"; (stage / "nested").mkdir(parents=True)   # only a subdir
    (stage / "~$lock.xlsx").write_bytes(b"x")        # + a lock file (excluded)
    check("directory-only / lock-only staging -> not promoted", a.promote_store(live, stage) is False)
    check("...the valid prior file set is preserved", sorted(p.name for p in live.iterdir()) == prior)


def test_b02_failed_rollback(tmp):                            # P2-B02
    print("P2-B02 a failed inline rollback RETAINS the journal so recovery can retry:")
    base = tmp / "b02a"; base.mkdir()
    _seed_store(base, "live")
    stage = base / "live.staging"; stage.mkdir(); (stage / "new.xlsx").write_bytes(b"new")
    # live->backup succeeds; staged->live and backup->live (restore) both fail.
    with _fail_rename_when(lambda s, t: t.name == "live"):
        ok = a.promote_store(base / "live", stage)
    check("promotion returns False", ok is False)
    check("...live is missing (death window) and a backup remains",
          not (base / "live").exists() and any(".bak-" in p.name for p in base.iterdir()))
    check("...the JOURNAL was RETAINED (not deleted) for retry",
          bool(list((base / ".promote").glob("*.json"))))
    a.recover_promotions(base, _OWN_ALL)                       # next launch
    check("the next-launch recovery RESTORES live from the retained backup",
          (base / "live").exists() and (base / "live" / "r1.xlsx").exists())

    print("P2-B02 a blocked first recovery retains the journal; the second succeeds:")
    base = tmp / "b02b"; base.mkdir()
    import json
    backup = _seed_store(base, "live.bak-tok")
    jdir = base / ".promote"; jdir.mkdir()
    (jdir / "tok.json").write_text(json.dumps(
        {"target": "live", "backup": "live.bak-tok", "staging": "live.staging", "token": "tok"}))
    with _fail_rename_when(lambda s, t: t.name == "live"):    # restore blocked
        a.recover_promotions(base, _OWN_ALL)
    check("first (blocked) sweep did NOT restore", not (base / "live").exists())
    check("...and RETAINED the journal + backup", backup.exists()
          and bool(list(jdir.glob("*.json"))))
    a.recover_promotions(base, _OWN_ALL)                       # unblocked second launch
    check("the second sweep restores live", (base / "live").exists()
          and (base / "live" / "r1.xlsx").exists())


def test_b04_traversal(tmp):                                  # P2-B04
    print("P2-B04 an untrusted journal cannot make recovery touch a path outside the store:")
    import json
    # The traversal target resolves to base/../victim (OUTSIDE the scanned store), and `live`
    # exists so the cleanup branch (`_rmtree(backup)` / `_rmtree(staging)`) would fire.
    cases = (("backup-traversal", {"target": "live", "backup": "../victim",
                                   "staging": "live.staging", "token": "t"}),
             ("staging-traversal", {"target": "live", "backup": "live.bak-t",
                                    "staging": "../victim", "token": "t"}),
             ("wrong-token-restore", {"target": "live", "backup": "live.bak-EVIL",
                                      "staging": "live.staging", "token": "t"}))
    for label, rec in cases:
        root = tmp / ("b04_" + label); root.mkdir()
        base = root / "store"; _seed_store(base, "live")        # target exists (cleanup path)
        victim = _seed_store(root, "victim")                    # base/../victim — out of store
        jdir = base / ".promote"; jdir.mkdir()
        (jdir / "t.json").write_text(json.dumps(rec))
        a.recover_promotions(base, _OWN_ALL)
        check(f"[{label}] the out-of-store victim dir is UNTOUCHED",
              victim.exists() and (victim / "r1.xlsx").exists())
        check(f"[{label}] the in-store live is preserved", (base / "live" / "r1.xlsx").exists())
        check(f"[{label}] the untrusted journal is dropped", not list(jdir.glob("*.json")))


def test_b02_invalid_target(tmp):                             # P2-B02 (round 2)
    print("P2-B02 recovery does NOT delete a valid backup when an INVALID placeholder is at live:")
    import json
    # (a) empty placeholder at live + valid backup -> displace placeholder, restore backup.
    base = tmp / "b02inv_a"; base.mkdir()
    (base / "live").mkdir()                          # empty placeholder — NOT a usable store
    backup = _seed_store(base, "live.bak-tok")       # the only real report data
    stage = _seed_store(base, "live.staging")
    jdir = base / ".promote"; jdir.mkdir()
    (jdir / "tok.json").write_text(json.dumps(
        {"target": "live", "backup": "live.bak-tok", "staging": "live.staging", "token": "tok"}))
    a.recover_promotions(base, _OWN_ALL)
    check("live now holds the restored report data (backup not deleted)",
          (base / "live" / "r1.xlsx").exists())
    check("...backup + staging + journal cleaned after a proven restore",
          not backup.exists() and not stage.exists() and not list(jdir.glob("*.json")))

    # (b) a placeholder with FOREIGN content + valid backup -> conflict: retain, touch nothing.
    base = tmp / "b02inv_b"; base.mkdir()
    ph = base / "live"; ph.mkdir(); (ph / "user_notes.txt").write_text("keep me")
    backup = _seed_store(base, "live.bak-tok")
    jdir = base / ".promote"; jdir.mkdir()
    (jdir / "tok.json").write_text(json.dumps(
        {"target": "live", "backup": "live.bak-tok", "staging": "live.staging", "token": "tok"}))
    a.recover_promotions(base, _OWN_ALL)
    check("a foreign-content placeholder is NOT destroyed", (ph / "user_notes.txt").exists())
    check("...the valid backup is RETAINED (conflict, not deleted)",
          backup.exists() and (backup / "r1.xlsx").exists())
    check("...the journal is RETAINED for manual resolution", bool(list(jdir.glob("*.json"))))


def test_b04_orphan_survives(tmp):                            # P2-B04 (round 2)
    print("P2-B04 an unrelated *.bak-* / *.staging directory with NO journal survives recovery:")
    root = tmp / "userdest"; root.mkdir()
    project = _seed_store(root, "Project")                    # a real same-prefix 'live' sibling
    photos = root / "Project.bak-family-photos"; photos.mkdir()
    (photos / "keep.txt").write_text("precious")              # unrelated user data, no journal
    other_stage = root / "Other.staging"; other_stage.mkdir(); (other_stage / "x.txt").write_text("x")
    a.recover_promotions(root, _OWN_ALL)
    check("the unrelated *.bak-* directory is UNTOUCHED", (photos / "keep.txt").exists())
    check("the unrelated *.staging directory is UNTOUCHED", other_stage.exists())
    check("the live sibling is untouched", (project / "r1.xlsx").exists())


def test_b06_first_promotion(tmp):                            # P2-B06
    print("P2-B06 a failed FIRST-ever promotion keeps the only copy + recovers it next launch:")
    base = tmp / "b06"; base.mkdir()
    live = base / "live"                                      # NO prior live
    stage = _seed_store(base, "live.staging")
    with _fail_rename_when(lambda s, t: t.name == "live"):    # the only rename (staging->live) fails
        ok = a.promote_store(live, stage)
    check("first promotion returns False", ok is False)
    check("...live was NOT created", not live.exists())
    check("...the staging (the ONLY completed copy) is RETAINED, not deleted",
          stage.exists() and (stage / "r1.xlsx").exists())
    check("...a journal was written so recovery can complete it", bool(list((base / ".promote").glob("*.json"))))
    a.recover_promotions(base, _OWN_ALL)                                # next launch
    check("recovery PROMOTES the surviving staging to live", (live / "r1.xlsx").exists())
    check("...staging consumed + journal cleared",
          not stage.exists() and not list((base / ".promote").glob("*.json")))


def test_r01_wrong_extension(tmp):                            # P2-R01 (round 2)
    print("P2-R01 staging with only a non-report file (.txt) is NOT promoted:")
    base = tmp / "r01ext"; base.mkdir()
    live = _seed_store(base, "live")
    prior = sorted(p.name for p in live.iterdir())
    stage = base / "live.staging"; stage.mkdir(); (stage / "notes.txt").write_text("not a report")
    check(".txt-only staging -> not promoted", a.promote_store(live, stage) is False)
    check("...the valid prior .xlsx store is preserved",
          sorted(p.name for p in live.iterdir()) == prior)
    # A real report artifact DOES promote (control).
    base = tmp / "r01ext_ok"; base.mkdir()
    live = _seed_store(base, "live")
    stage = base / "live.staging"; stage.mkdir(); (stage / "r9.xlsx").write_bytes(b"real")
    check("control: a .xlsx staging DOES promote", a.promote_store(live, stage) is True)


def test_r04_malformed_sheet(tmp):                            # P2-R04
    print("P2-R04 a malformed worksheet is rejected, prior preserved, NO locked temp residue:")
    d = tmp / "r04"; d.mkdir()
    final = d / "cmp.xlsx"
    _xlsx(final, sheet="Comparison"); prior = final.read_bytes()

    def produce_bad_sheet(t):
        _workbook_with_bad_sheet(Path(t))
        return ConsolidateResult(status="ok", output_path=str(t))
    res = a.commit_workbook(final, produce_bad_sheet, expect_sheet="Comparison")
    check("a corrupt-worksheet workbook -> error (rejected)", res.status == "error")
    check("...the valid prior workbook is preserved", final.read_bytes() == prior)
    check("...NO temp residue remains (handle released, temp removed)", _no_temp_leak(d))


def test_a02_revert_race(tmp):                                # P2-A02 (round 2)
    print("P2-A02 a detected build race REMOVES the stale sidecar (no false-fresh after replace):")
    store = tmp / "a02r"; store.mkdir()
    (store / "r1.xlsx").write_bytes(b"identity-B")
    consolidated = tmp / "a02r_combined.xlsx"; _xlsx(consolidated)
    a.write_consolidated_fingerprint(consolidated, store)     # sidecar certifies identity B
    check("freshly recorded -> fresh", a.consolidated_fresh(consolidated, store))
    # A rebuild STARTED from identity A; the producer already replaced the workbook; the inputs
    # then reverted EXACTLY to B before publication. The old B sidecar must NOT certify it.
    ok = a.write_consolidated_fingerprint(consolidated, store, built_from="v1:0:identityA")
    check("the race is detected (returns False)", ok is False)
    check("...the stale sidecar was REMOVED so the workbook reads stale + rebuilds",
          a.consolidated_fresh(consolidated, store) is False)


def test_b07_two_journals_generation_aware(tmp):              # P2-B07
    print("P2-B07 two same-target journals: recovery keeps the NEWEST last-good (any order):")
    import json
    # The reachable interrupted two-generation state: a V1->V2 promotion whose backup cleanup was
    # locked left an OLDER journal (gen=1, backup=V1); a later V2->V3 promotion was interrupted after
    # live->backup (NEWER journal gen=2, backup=V2=the true last-good, shared det staging=V3, live
    # ABSENT). Vary which journal NAME the older/newer get, so the durable `gen` (not the filesystem
    # order) is what guarantees newest-first in BOTH encounter orders.
    for older_name, newer_name in (("a.json", "z.json"), ("z.json", "a.json")):
        tag = f"{older_name}-{newer_name}"
        root = tmp / ("b07_" + tag); root.mkdir()
        store_root = root / "ssor-prod"; store_root.mkdir()
        jdir = store_root / ".promote"; jdir.mkdir()

        def _store(name, content):
            d = store_root / name; d.mkdir()
            (d / "r.xlsx").write_text(content)
            return d
        bak_v1 = _store("highway_log.bak-t1", "V1")
        _store("highway_log.bak-t2", "V2")                   # the NEWER backup = the true last-good
        staging_v3 = _store("highway_log.staging", "V3")     # shared det staging (newer, interrupted)
        (jdir / older_name).write_text(json.dumps(
            {"target": "highway_log", "backup": "highway_log.bak-t1",
             "staging": "highway_log.staging", "token": "t1", "gen": 1}))
        (jdir / newer_name).write_text(json.dumps(
            {"target": "highway_log", "backup": "highway_log.bak-t2",
             "staging": "highway_log.staging", "token": "t2", "gen": 2}))

        def _is_owned(sr, target):
            return sr.name == "ssor-prod" and (target is None or target == "highway_log")
        a.recover_promotions(root, _is_owned)

        live = store_root / "highway_log"
        check(f"[{tag}] live holds the NEWEST proven last-good V2",
              live.is_dir() and (live / "r.xlsx").read_text() == "V2")
        check(f"[{tag}] the older generation V1 backup is gone (never restored over V2)",
              not bak_v1.exists())
        check(f"[{tag}] the un-promoted V3 staging is gone", not staging_v3.exists())
        check(f"[{tag}] both journals are cleared", not list(jdir.glob("*.json")))


def test_b07_durable_generation(tmp):                         # P2-B07 (round 6)
    print("P2-B07 the transaction generation is DURABLE (derived from journals, not wall-clock):")
    import json
    base = tmp / "gen_unit"; base.mkdir()
    jdir = base / ".promote"; jdir.mkdir()
    check("no prior same-target journal -> generation 1", a._next_generation(jdir, "live") == 1)
    (jdir / "j1.json").write_text(json.dumps(
        {"target": "live", "backup": "live.bak-x", "staging": "live.staging", "token": "x", "gen": 7}))
    check("a prior gen=7 same-target journal -> next is 8 (max+1)", a._next_generation(jdir, "live") == 8)
    (jdir / "j2.json").write_text(json.dumps(
        {"target": "elsewhere", "backup": "elsewhere.bak-y", "staging": "elsewhere.staging",
         "token": "y", "gen": 99}))
    check("a DIFFERENT target's higher gen is ignored", a._next_generation(jdir, "live") == 8)

    # Two REAL same-target promotions: the 2nd's journal generation is strictly GREATER, derived
    # from the on-disk journals — so it NEVER regresses even if the wall clock moves backward
    # between them (the round-6 repro). The 1st journal is retained (its backup cleanup is locked),
    # so both coexist for the comparison.
    live = tmp / "gen_real" / "ssor-prod" / "highway_log"; live.mkdir(parents=True)
    (live / "r.xlsx").write_text("V1")
    pjdir = live.parent / ".promote"
    orig_rmtree = a._rmtree

    def _block_bak(p, expected_identity=None):
        if a._BAK_INFIX in Path(p).name:
            return                                           # the renamed-aside backup is "locked"
        return orig_rmtree(p, expected_identity)

    def _promote(content):
        stage = live.with_name("highway_log.staging"); stage.mkdir(exist_ok=True)
        (stage / "r.xlsx").write_text(content)
        with _patch(a, "_rmtree", _block_bak):
            a.promote_store(live, stage)
    _promote("V2")                                           # gen 1, journal retained (backup locked)
    _promote("V3")                                           # gen 2, journal retained
    gens = sorted(a._journal_gen(p) for p in pjdir.glob("*.json"))
    check("two same-target journals coexist (both backups locked)", len(gens) == 2)
    check("their generations are strictly increasing 1,2 (a small counter, not a timestamp)",
          gens == [1, 2])


def test_b04_unowned_journal(tmp):                            # P2-B04 (rounds 3 + 4)
    print("P2-B04 ownership is LOCATION-based: a nested valid-NAME store is rejected before any read:")
    import json
    root = tmp / "userroot"; root.mkdir()

    # A location-only mechanics predicate: direct-child + known names. The updater's
    # stronger current-marker and captured-identity policy has its own integration
    # regression below.
    owned_roots = {"ssor-prod", "ars-test"}
    owned_targets = {"highway_log", "ramp_summary"}

    def _is_owned(store_root, target):
        if store_root.parent.resolve() != root.resolve():
            return False
        if store_root.name not in owned_roots:
            return False
        return target is None or target in owned_targets

    # (1) Round-4 repro: a valid `<src>-<env>` NAME nested one level too deep (wrong LOCATION) —
    # <root>/UnrelatedProject/ssor-prod/.promote with a SHAPE-VALID journal for a real target.
    nested_parent = root / "UnrelatedProject" / "ssor-prod"
    target = _seed_store(nested_parent, "ramp_summary")     # a usable store (r1.xlsx)
    backup = _seed_store(nested_parent, "ramp_summary.bak-tok")   # unrelated user data
    stage = _seed_store(nested_parent, "ramp_summary.staging")
    njdir = nested_parent / ".promote"; njdir.mkdir()
    njournal = njdir / "tok.json"
    njournal.write_text(json.dumps(
        {"target": "ramp_summary", "backup": "ramp_summary.bak-tok",
         "staging": "ramp_summary.staging", "token": "tok"}))
    # (2) A nested MALFORMED journal at a wrong location must ALSO be left untouched.
    mal_parent = root / "Other" / "ars-test"; mal_parent.mkdir(parents=True)
    mjdir = mal_parent / ".promote"; mjdir.mkdir()
    mjournal = mjdir / "bad.json"; mjournal.write_text("{ not json")
    # (3) Owned direct-child control: <root>/ssor-prod/.promote (live missing) -> restores.
    owned_root = root / "ssor-prod"; owned_root.mkdir()
    _seed_store(owned_root, "highway_log.bak-t2")           # backup; live (target) is missing
    ojdir = owned_root / ".promote"; ojdir.mkdir()
    (ojdir / "t2.json").write_text(json.dumps(
        {"target": "highway_log", "backup": "highway_log.bak-t2",
         "staging": "highway_log.staging", "token": "t2"}))

    a.recover_promotions(root, _is_owned)

    check("the nested valid-name target is UNTOUCHED (wrong location)", (target / "r1.xlsx").exists())
    check("the nested backup is UNTOUCHED", (backup / "r1.xlsx").exists())
    check("the nested staging is UNTOUCHED", (stage / "r1.xlsx").exists())
    check("the nested shape-valid journal is left in place (not ours)", njournal.exists())
    check("the nested MALFORMED journal is NOT deleted (outside an owned location)", mjournal.exists())
    check("control: the OWNED direct-child journal restored its target from backup",
          (owned_root / "highway_log" / "r1.xlsx").exists())


def test_r05_promote_cleanup(tmp):                            # P2-R05 (round 4: the promote_store path)
    print("P2-R05 promote_store retains the journal when its OWN residue cleanup fails:")
    orig_rmtree = a._rmtree

    # (a) successful promotion (prior exists) with a LOCKED backup -> journal RETAINED.
    base = tmp / "r05p_a"; base.mkdir()
    live = _seed_store(base, "live")                         # a prior live store
    stage = base / "live.staging"; stage.mkdir(); (stage / "r9.xlsx").write_bytes(b"new")

    def _block_bak(p, expected_identity=None):
        if a._BAK_INFIX in Path(p).name:
            return                                           # the renamed-aside backup is "locked"
        return orig_rmtree(p, expected_identity)
    with _patch(a, "_rmtree", _block_bak):
        ok = a.promote_store(live, stage)
    check("the promotion still succeeds (new data is live)",
          ok is True and (live / "r9.xlsx").exists())
    check("...the un-removable backup survived", any(a._BAK_INFIX in p.name for p in base.iterdir()))
    check("...the journal was RETAINED for next-launch cleanup",
          bool(list((base / ".promote").glob("*.json"))))
    a.recover_promotions(base, _OWN_ALL)                     # a later, unblocked sweep cleans it
    check("a later recovery removes the leftover backup",
          not any(a._BAK_INFIX in p.name for p in base.iterdir()))
    check("...and then drops the journal", not list((base / ".promote").glob("*.json")))

    # (b) inline restore (staging->live fails, restore succeeds) with a LOCKED staging -> RETAINED.
    base = tmp / "r05p_b"; base.mkdir()
    live = _seed_store(base, "live")
    stage = base / "live.staging"; stage.mkdir(); (stage / "r9.xlsx").write_bytes(b"new")

    def _fail_stage_to_live(s, t):                           # only staging->live fails (force restore)
        return t.name == "live" and s.name.endswith(a._STAGING_SUFFIX)

    def _block_stage_rm(p, expected_identity=None):
        if Path(p).name.endswith(a._STAGING_SUFFIX):
            return                                           # the residual staging is "locked"
        return orig_rmtree(p, expected_identity)
    with _fail_rename_when(_fail_stage_to_live), _patch(a, "_rmtree", _block_stage_rm):
        ok = a.promote_store(live, stage)
    check("the inline restore kept last-good live", (live / "r1.xlsx").exists() and ok is False)
    check("...the un-removable staging survived", stage.exists())
    check("...the journal was RETAINED (residue cleanup incomplete)",
          bool(list((base / ".promote").glob("*.json"))))


def test_a02_dual_failure_quarantine(tmp):                    # P2-A02 (round 4)
    print("P2-A02 if BOTH sidecar removal AND sentinel write fail, the workbook is quarantined:")
    store = tmp / "a02d"; store.mkdir()
    (store / "r1.xlsx").write_bytes(b"identity-B")
    consolidated = tmp / "a02d_combined.xlsx"; _xlsx(consolidated)
    a.write_consolidated_fingerprint(consolidated, store)    # sidecar certifies identity B
    check("freshly recorded -> fresh", a.consolidated_fresh(consolidated, store))
    orig_unlink = a._silent_unlink

    def _block_sidecar(p):
        if str(p).endswith(a._FP_SUFFIX):
            return False                                     # the sidecar is "locked" (can't unlink)
        return orig_unlink(p)
    with _patch(a, "_silent_unlink", _block_sidecar), \
         _patch(a, "_write_fp_sentinel", lambda _p: False):  # AND the sentinel write fails
        ok = a.write_consolidated_fingerprint(consolidated, store, built_from="v1:0:identityA")
    check("the race still returns False", ok is False)
    check("...the replaced workbook was QUARANTINED (canonical path now missing)",
          not consolidated.exists())
    check("...so consolidated_fresh reads STALE despite the un-removable sidecar",
          a.consolidated_fresh(consolidated, store) is False)


def test_r05_cleanup_failure_retains_journal(tmp):            # P2-R05
    print("P2-R05 a failed residue cleanup RETAINS the journal so a later sweep retries:")
    import json
    orig_rmtree = a._rmtree

    # (a) usable target + a backup whose removal fails -> journal retained, retried later.
    base = tmp / "r05a"; base.mkdir()
    _seed_store(base, "live")                                # a usable canonical target
    backup = _seed_store(base, "live.bak-tok")
    jdir = base / ".promote"; jdir.mkdir()
    jpath = jdir / "tok.json"
    jpath.write_text(json.dumps(
        {"target": "live", "backup": "live.bak-tok", "staging": "live.staging", "token": "tok"}))

    def _block_backup(p, expected_identity=None):
        if Path(p).name == "live.bak-tok":
            return                                           # simulate a locked backup (not removed)
        return orig_rmtree(p, expected_identity)
    with _patch(a, "_rmtree", _block_backup):
        a.recover_promotions(base, _OWN_ALL)
    check("the un-removable backup survived", backup.exists())
    check("...and the journal was RETAINED (cleanup incomplete)", jpath.exists())
    a.recover_promotions(base, _OWN_ALL)                     # a later, unblocked sweep
    check("a later sweep removes the backup", not backup.exists())
    check("...and then drops the journal", not jpath.exists())

    # (b) after a RESTORE, a residual staging that can't be removed retains the journal.
    base = tmp / "r05b"; base.mkdir()
    backup = _seed_store(base, "live.bak-t")                 # target missing -> restore from backup
    stage = _seed_store(base, "live.staging")
    jdir = base / ".promote"; jdir.mkdir()
    jpath = jdir / "t.json"
    jpath.write_text(json.dumps(
        {"target": "live", "backup": "live.bak-t", "staging": "live.staging", "token": "t"}))

    def _block_staging(p, expected_identity=None):
        if Path(p).name == "live.staging":
            return                                           # locked staging
        return orig_rmtree(p, expected_identity)
    with _patch(a, "_rmtree", _block_staging):
        a.recover_promotions(base, _OWN_ALL)
    check("the restore succeeded (live present)", (base / "live" / "r1.xlsx").exists())
    check("...but the un-removable staging RETAINS the journal", jpath.exists() and stage.exists())


def test_a02_sidecar_unremovable(tmp):                        # P2-A02 (round 3)
    print("P2-A02 an un-removable stale sidecar is overwritten with a non-matching sentinel:")
    store = tmp / "a02s"; store.mkdir()
    (store / "r1.xlsx").write_bytes(b"identity-B")
    consolidated = tmp / "a02s_combined.xlsx"; _xlsx(consolidated)
    a.write_consolidated_fingerprint(consolidated, store)    # sidecar certifies identity B
    check("freshly recorded -> fresh", a.consolidated_fresh(consolidated, store))
    orig_unlink = a._silent_unlink

    def _block_sidecar(p):
        if str(p).endswith(a._FP_SUFFIX):
            return False                                     # simulate a locked sidecar (not removed)
        return orig_unlink(p)
    with _patch(a, "_silent_unlink", _block_sidecar):
        ok = a.write_consolidated_fingerprint(consolidated, store, built_from="v1:0:identityA")
    check("the race still returns False", ok is False)
    check("...and the workbook reads STALE despite the un-removable sidecar (sentinel written)",
          a.consolidated_fresh(consolidated, store) is False)


def test_a02_build_race(tmp):                                 # P2-A02
    print("P2-A02 a fingerprint is NOT certified fresh when inputs changed during the build:")
    store = tmp / "a02"; store.mkdir()
    (store / "r1.xlsx").write_bytes(b"a" * 10)
    consolidated = tmp / "a02_combined.xlsx"; _xlsx(consolidated)
    fp_before = a.fingerprint(store)
    (store / "r2.xlsx").write_bytes(b"b" * 10)    # an input changed AFTER fp_before was captured
    ok = a.write_consolidated_fingerprint(consolidated, store, built_from=fp_before)
    check("the sidecar is NOT written (the build raced an input change)", ok is False)
    check("...so the workbook reads stale and rebuilds next reuse",
          a.consolidated_fresh(consolidated, store) is False)
    ok2 = a.write_consolidated_fingerprint(consolidated, store, built_from=a.fingerprint(store))
    check("no race (before == after) -> the sidecar IS written",
          ok2 is True and a.consolidated_fresh(consolidated, store))


def test_b03_updater_recovers_batch_dest(tmp):                # P2-B03
    print("P2-B03 startup recovery sweeps the configured batch destination, not just OUTPUT_ROOT:")
    import updater
    import paths
    import settings
    swept = []
    custom = tmp / "custom_batch_dest"
    with _patch(a, "recover_promotions", lambda r, _own: swept.append(Path(r))), \
         _patch(settings, "get_batch_dest", lambda: str(custom)):
        updater._recover_store_promotions()
    swept_resolved = {p.resolve() if p.exists() else p for p in swept}
    check("OUTPUT_ROOT is swept", Path(paths.OUTPUT_ROOT) in swept
          or Path(paths.OUTPUT_ROOT).resolve() in swept_resolved)
    check("the CUSTOM batch destination (outside OUTPUT_ROOT) is also swept", custom in swept)

    # Deduplication: when the batch dest == OUTPUT_ROOT, the sweep runs once.
    swept.clear()
    with _patch(a, "recover_promotions", lambda r, _own: swept.append(Path(r))), \
         _patch(settings, "get_batch_dest", lambda: str(paths.OUTPUT_ROOT)):
        updater._recover_store_promotions()
    check("the batch-dest == OUTPUT_ROOT case sweeps exactly once (deduped)", len(swept) == 1)


def test_promote_ownership_lease(tmp):                         # CMP-AUD-090
    print("CMP-AUD-090 store promotion rechecks its leased root before each mutation:")
    import owned_dir

    root = owned_dir.ensure_owned_dir(tmp / "leased-promotion", kind="store")
    lease = owned_dir.require_owned_dir_lease(root, kind="store")
    live = _seed_store(root, "ramp_summary")
    staged = _seed_store(root, "ramp_summary.staging")
    moved = tmp / "leased-promotion-original-moved"

    replacement_source = owned_dir.ensure_owned_dir(
        tmp / "different-promotion-store", kind="store")
    _seed_store(replacement_source, "ramp_summary")
    _seed_store(replacement_source, "ramp_summary.staging")

    def _tree_bytes(tree):
        return {
            str(p.relative_to(tree)): (("dir", None) if p.is_dir()
                                      else ("file", p.read_bytes()))
            for p in sorted(Path(tree).rglob("*"))
        }

    before_replacement = _tree_bytes(replacement_source)
    calls = 0

    def _guard():
        nonlocal calls
        calls += 1
        # Calls 1/2 guard journal-dir creation and journal write. Swap roots
        # immediately before the live->backup rename boundary.
        if calls == 3:
            root.rename(moved)
            replacement_source.rename(root)
        return lease.is_current()

    promoted = a.promote_store(live, staged, guard=_guard)
    check("promotion aborts when the root identity changes after journaling",
          calls >= 3 and promoted is False)
    check("the different current-owned replacement remains byte-exact",
          _tree_bytes(root) == before_replacement)
    check("the original live/staging data remains under its original identity",
          (moved / "ramp_summary" / "r1.xlsx").exists()
          and (moved / "ramp_summary.staging" / "r1.xlsx").exists())


def _directory_link(link, target):
    """Create a directory junction/symlink for reparse-boundary regressions."""
    try:
        if os.name == "nt":
            made = subprocess.run(
                ["cmd", "/c", "mklink", "/J", str(link), str(target)],
                capture_output=True, text=True, timeout=10)
            return made.returncode == 0
        os.symlink(target, link, target_is_directory=True)
        return True
    except (OSError, subprocess.SubprocessError):
        return False


def test_promote_reparse_journal_dir(tmp):                     # CMP-AUD-090
    print("CMP-AUD-090 promotion/recovery never follows a reparse .promote directory:")
    external = tmp / "external-journals"
    external.mkdir()

    promote_root = tmp / "reparse-promote-store"
    promote_root.mkdir()
    live = _seed_store(promote_root, "ramp_summary")
    staged = _seed_store(promote_root, "ramp_summary.staging")
    planted = external / "attack.json"
    planted.write_bytes(b"external personal journal")
    if not _directory_link(promote_root / ".promote", external):
        print("  [SKIP] no directory junction/symlink could be created")
        return

    check("generation scan rejects a preplanted reparse .promote before listing it",
          a._next_generation(promote_root / ".promote", "ramp_summary") is None)
    with _patch(a, "_new_token", lambda: "attack"):
        promoted = a.promote_store(live, staged)
    check("promotion rejects a preplanted reparse .promote root",
          promoted is False)
    check("promotion neither rewrites nor deletes external JSON",
          planted.exists() and planted.read_bytes() == b"external personal journal")
    check("rejected promotion keeps the canonical live store",
          (live / "r1.xlsx").exists())

    recovery_root = tmp / "reparse-recovery-root"
    store = recovery_root / "ssor-prod"
    store.mkdir(parents=True)
    external_recovery = tmp / "external-recovery-journals"
    external_recovery.mkdir()
    external_journal = external_recovery / "planted.json"
    external_journal.write_bytes(b"not app data")
    if not _directory_link(store / ".promote", external_recovery):
        print("  [SKIP] recovery junction/symlink could not be created")
        return
    a.recover_promotions(recovery_root, _OWN_ALL)
    check("recovery leaves external JSON byte-exact behind reparse .promote",
          external_journal.exists() and external_journal.read_bytes() == b"not app data")


def test_journal_boundary_replacements(tmp):                  # CMP-AUD-090
    print("CMP-AUD-090 journal scans/reads/cleanup stay bound to captured entries:")
    import json

    # Generation scan: the ordinary directory captured by the caller is moved
    # aside and a different ordinary directory lands at the same pathname.
    base = tmp / "journal-generation-race"; base.mkdir()
    jdir = base / ".promote"; jdir.mkdir()
    (jdir / "old.json").write_text(json.dumps({
        "target": "live", "backup": "live.bak-old",
        "staging": "live.staging", "token": "old", "gen": 4,
    }), encoding="utf-8")
    captured_dir = a._plain_journal_dir_identity(jdir)
    moved = base / ".promote-captured"; jdir.rename(moved)
    replacement = base / ".promote-replacement"; replacement.mkdir()
    attack = replacement / "attack.json"
    attack.write_bytes(b'external replacement bytes')
    replacement.rename(jdir)
    attack = jdir / "attack.json"
    before = attack.read_bytes()
    check("generation scan rejects a same-path replacement directory",
          a._next_generation(jdir, "live", captured_dir) is None)
    check("...and neither reads as authority nor mutates replacement JSON",
          attack.read_bytes() == before)

    # Bound read: inject a directory replacement at the final os.open boundary.
    # The descriptor points at the replacement file, but its identity differs
    # from the captured regular journal, so fdopen/JSON consumption never occurs.
    base = tmp / "journal-read-race"; base.mkdir()
    jdir = base / ".promote"; jdir.mkdir()
    journal = jdir / "tok.json"
    journal.write_text(json.dumps({
        "target": "live", "backup": "live.bak-tok",
        "staging": "live.staging", "token": "tok", "gen": 1,
    }), encoding="utf-8")
    captured_dir = a._plain_journal_dir_identity(jdir)
    captured_journal = a._plain_journal_identity(jdir, captured_dir, journal)
    replacement = base / ".promote-other"; replacement.mkdir()
    replacement_journal = replacement / "tok.json"
    replacement_journal.write_bytes(b"private replacement contents")
    moved = base / ".promote-original"
    original_open, original_fdopen = a.os.open, a.os.fdopen
    consumed = False

    def _racing_open(path, flags, *args, **kwargs):
        jdir.rename(moved)
        replacement.rename(jdir)
        return original_open(path, flags, *args, **kwargs)

    def _tracked_fdopen(*args, **kwargs):
        nonlocal consumed
        consumed = True
        return original_fdopen(*args, **kwargs)

    with _patch(a.os, "open", _racing_open), _patch(a.os, "fdopen", _tracked_fdopen):
        value = a._read_bound_journal(
            jdir, captured_dir, journal, captured_journal)
    check("a journal swapped at open is rejected before any JSON bytes are consumed",
          value is a._JOURNAL_UNSAFE and not consumed)
    check("...the replacement journal remains byte-exact",
          (jdir / "tok.json").read_bytes() == b"private replacement contents")

    # Root cleanup: swap a captured ordinary residue for unrelated content at
    # the _rmtree call boundary. The expected identity travels into _rmtree.
    base = tmp / "residue-delete-race"; base.mkdir()
    residue = base / "live.bak-tok"; residue.mkdir()
    (residue / "old.xlsx").write_bytes(b"old app residue")
    moved = base / "captured-residue"
    replacement = base / "unrelated-replacement"; replacement.mkdir()
    (replacement / "personal.txt").write_bytes(b"keep me")
    before = (replacement / "personal.txt").read_bytes()
    original_rmtree = a._rmtree

    def _swap_then_remove(path, expected_identity=None):
        residue.rename(moved)
        replacement.rename(residue)
        return original_rmtree(path, expected_identity)

    with _patch(a, "_rmtree", _swap_then_remove):
        removed = a._rmtree_gone(residue)
    check("cleanup rejects a same-path replacement after identity capture",
          removed is False)
    check("...both the unrelated replacement and original residue survive",
          (residue / "personal.txt").read_bytes() == before
          and (moved / "old.xlsx").read_bytes() == b"old app residue")

    # Absence is an identity state too: a residue that appears after the cleanup
    # snapshot is foreign, not newly-authorized app residue.
    base = tmp / "late-residue"; base.mkdir()
    journal = base / "journal.json"; journal.write_text("{}", encoding="utf-8")
    late = base / "live.staging"; late.mkdir()
    (late / "personal.txt").write_bytes(b"late foreign data")
    finalized = a._finalize_journal(
        journal, late, residue_identities={late: a._ABSENT_ENTRY})
    check("cleanup does not delete a residue that appeared after absence was captured",
          finalized is False and journal.exists()
          and (late / "personal.txt").read_bytes() == b"late foreign data")


def test_reparse_journal_entries_and_broken_residue(tmp):     # CMP-AUD-090
    print("CMP-AUD-090 reparse journal entries/residue are retained, never followed:")
    import json

    # A `.json` directory junction is not a regular journal entry. Recovery
    # must neither read through it nor delete the external directory.
    base = tmp / "reparse-journal-entry"; base.mkdir()
    _seed_store(base, "live")
    jdir = base / ".promote"; jdir.mkdir()
    external = tmp / "external-journal-entry"; external.mkdir()
    victim = external / "private.txt"; victim.write_bytes(b"private")
    link = jdir / "attack.json"
    if _directory_link(link, external):
        a.recover_promotions(base, _OWN_ALL)
        check("a reparse .json entry is left in place",
              a._entry_present_lstat(link))
        check("...and its external target remains byte-exact",
              victim.read_bytes() == b"private")
    else:
        print("  [SKIP] reparse journal-entry probe unavailable")

    # A broken junction reports exists()==False, but lstat still sees residue.
    # The journal must remain so cleanup can never be falsely certified.
    base = tmp / "broken-reparse-residue"; base.mkdir()
    _seed_store(base, "live")
    jdir = base / ".promote"; jdir.mkdir()
    journal = jdir / "tok.json"
    journal.write_text(json.dumps({
        "target": "live", "backup": "live.bak-tok",
        "staging": "live.staging", "token": "tok",
    }), encoding="utf-8")
    target = tmp / "eventually-missing-target"; target.mkdir()
    broken = base / "live.bak-tok"
    if _directory_link(broken, target):
        target.rmdir()
        check("the probe is a present lstat entry even when its target is gone",
              a._entry_present_lstat(broken))
        a.recover_promotions(base, _OWN_ALL)
        check("a broken reparse residue is not misclassified as cleaned",
              a._entry_present_lstat(broken) and journal.exists())
    else:
        print("  [SKIP] broken-reparse residue probe unavailable")


def test_b04_updater_requires_creation_claim(tmp):            # CMP-AUD-090 / P2-B04
    print("CMP-AUD-090 startup recovery requires current ownership in a user destination:")
    import json
    import owned_dir
    import paths
    import settings
    import updater

    app_output = tmp / "private-output"
    custom = tmp / "user-destination"
    app_output.mkdir()
    custom.mkdir()

    def _journal(store, target, token, *, live=True):
        if live:
            _seed_store(store, target)
        _seed_store(store, f"{target}.bak-{token}")
        _seed_store(store, f"{target}.staging")
        jdir = store / ".promote"
        jdir.mkdir()
        (jdir / f"{token}.json").write_text(json.dumps({
            "target": target,
            "backup": f"{target}.bak-{token}",
            "staging": f"{target}.staging",
            "token": token,
        }), encoding="utf-8")

    def _tree_bytes(root):
        result = {}
        for p in sorted(Path(root).rglob("*")):
            rel = str(p.relative_to(root))
            result[rel] = ("dir", None) if p.is_dir() else ("file", p.read_bytes())
        return result

    # Both direct children have valid app-shaped names and journals, but neither
    # has a current create-and-mark claim. The legacy marker is specifically the
    # pre-Phase-1 format that Reset already treats as untrusted.
    unowned = custom / "ssor-prod"
    unowned.mkdir()
    _journal(unowned, "ramp_summary", "u")
    legacy = custom / "ars-test"
    legacy.mkdir()
    (legacy / owned_dir.OWNER_MARKER).write_text(json.dumps({
        "app": "TSMIS Reports Exporter", "schema": 1, "kind": "store",
    }), encoding="utf-8")
    _journal(legacy, "highway_log", "l")
    before_unowned = _tree_bytes(unowned)
    before_legacy = _tree_bytes(legacy)

    # Positive controls: a current purpose-bound store in the same USER root is
    # recoverable, and the exact app-private OUTPUT_ROOT retains its deliberate
    # legacy name-based recovery policy for pre-marker interrupted stores.
    current = owned_dir.ensure_owned_dir(custom / "ars-prod", kind="store")
    _journal(current, "highway_log", "c", live=False)
    private_legacy = app_output / "ssor-test"
    private_legacy.mkdir()
    _journal(private_legacy, "ramp_summary", "p", live=False)

    with _patch(paths, "OUTPUT_ROOT", app_output), \
         _patch(settings, "get_batch_dest", lambda: str(custom)):
        updater._recover_store_promotions()

    check("unowned user-destination store remains byte-exact",
          _tree_bytes(unowned) == before_unowned)
    check("legacy-marked user-destination store remains byte-exact",
          _tree_bytes(legacy) == before_legacy)
    check("current purpose-bound user store is recovered",
          (current / "highway_log" / "r1.xlsx").exists())
    check("exact app-private OUTPUT_ROOT keeps legacy recovery compatibility",
          (private_legacy / "ramp_summary" / "r1.xlsx").exists())

    # Replacement race: even another genuinely app-created store cannot inherit
    # this recovery pass's authority merely by moving onto the captured pathname.
    raced = owned_dir.ensure_owned_dir(custom / "ssor-dev", kind="store")
    _journal(raced, "ramp_summary", "r")
    moved_raced = custom / "ssor-dev-original-moved"
    replacement_source = owned_dir.ensure_owned_dir(
        tmp / "different-owned-store", kind="store")
    _journal(replacement_source, "ramp_summary", "r")
    before_replacement = _tree_bytes(replacement_source)
    original_usable = a.is_usable_store
    swapped = False

    def _replace_after_location_claim(path):
        nonlocal swapped
        usable = original_usable(path)
        if not swapped and Path(path) == raced / "ramp_summary":
            swapped = True
            raced.rename(moved_raced)
            replacement_source.rename(raced)
        return usable

    with _patch(paths, "OUTPUT_ROOT", app_output), \
         _patch(settings, "get_batch_dest", lambda: str(custom)), \
         _patch(a, "is_usable_store", _replace_after_location_claim):
        updater._recover_store_promotions()

    check("a different current-owned directory cannot inherit a captured recovery lease",
          swapped and _tree_bytes(raced) == before_replacement)
    check("the originally leased store survives at its raced-aside identity",
          (moved_raced / "ramp_summary" / "r1.xlsx").exists())


def main():
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        test_atomic_save(tmp)
        test_commit_single(tmp)
        test_commit_twin(tmp)
        test_comparison_artifact_generation(tmp)
        test_comparison_generation_failure_paths(tmp)
        test_commit_target_guard(tmp)
        test_output_source_alias_guard(tmp)
        test_b05_malformed(tmp)
        test_r02_no_temp_leak(tmp)
        test_r03_formulas_truthful(tmp)
        test_fingerprint(tmp)
        test_consolidated_fresh(tmp)
        test_fingerprint_commit_guard(tmp)
        test_promote_validate(tmp)
        test_promote_ownership_lease(tmp)
        test_promote_reparse_journal_dir(tmp)
        test_journal_boundary_replacements(tmp)
        test_reparse_journal_entries_and_broken_residue(tmp)
        test_r01_directory_staging(tmp)
        test_r01_wrong_extension(tmp)
        test_r04_malformed_sheet(tmp)
        test_b02_failed_rollback(tmp)
        test_b02_invalid_target(tmp)
        test_b04_traversal(tmp)
        test_b04_orphan_survives(tmp)
        test_b04_unowned_journal(tmp)
        test_b06_first_promotion(tmp)
        test_b07_two_journals_generation_aware(tmp)
        test_b07_durable_generation(tmp)
        test_r05_cleanup_failure_retains_journal(tmp)
        test_r05_promote_cleanup(tmp)
        test_a02_build_race(tmp)
        test_a02_revert_race(tmp)
        test_a02_sidecar_unremovable(tmp)
        test_a02_dual_failure_quarantine(tmp)
        test_b03_updater_recovers_batch_dest(tmp)
        test_b04_updater_requires_creation_claim(tmp)
        test_recovery(tmp)
    print()
    if _failures:
        print(f"FAILED {len(_failures)} check(s): {_failures}")
        return 1
    print("All artifact_store checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
