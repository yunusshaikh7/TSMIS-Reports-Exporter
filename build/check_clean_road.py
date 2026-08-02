"""Clean Road CA HIGHWAYS (v0.29.0) — the ArcGIS layer substrate, the overlay
consolidator, the TSN normalizer, and the ArcGIS-vs-TSN comparator, all on a
SYNTHETIC mini-library (hermetic: no real data, no network).

Covers:
  * the dialect normalizers (labels export <-> bundle codes <-> TASAS codes);
  * LRS as-of algebra + integer micro-postmile round-trips;
  * stream_layer's name-keyed reads, optional columns, and the INDEX
    row-count gate (truncation refuses; the measured healthy over-count race
    passes);
  * the consolidator end-to-end on a tiny library: base/R/L rows, the X
    coverage-gap row, city cuts, the ADT profile family, TOLL/FOREST mux,
    point attachments, cross-county splitting, the 74-column header, the
    Provenance sheet (every column tiered), the build marker, and the
    missing-layer / truncated-layer refusals;
  * the TSN normalizer (verbatim projection + CMP-AUD-037 marker) and the
    comparator's role gates (the ArcGIS side REQUIRES the build marker, the
    TSN side REJECTS it) + a real mode="both" comparison where a CONTEXT
    column's one-sided values are never counted as differences.

Run from the repo root:
    build\\.venv\\Scripts\\python.exe build\\check_clean_road.py
"""
import json
import sys
import tempfile
import textwrap
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

import city_codes                       # noqa: E402
import clean_highway_columns as chc     # noqa: E402
import clean_road_layers as crl         # noqa: E402
import compare_clean_highway_tsn as cht  # noqa: E402
import consolidate_clean_highway as cch  # noqa: E402
import tsn_load_clean_road as tlc       # noqa: E402
from events import Events               # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


# --------------------------------------------------------------------------- #
# fixtures
# --------------------------------------------------------------------------- #
ASOF = "2025-09-08"
_LIVE = {"LRSFromDate": datetime(2020, 1, 1), "LRSToDate": None}
_DEAD = {"LRSFromDate": datetime(2010, 1, 1), "LRSToDate": datetime(2015, 1, 1)}
_FUTURE = {"LRSFromDate": datetime(2026, 1, 1), "LRSToDate": None}


def _span(county, b, e, od, attrs, align="Right", prefix=".", route="1",
          district="District 12", county2=None, prefix2=None, life=None,
          item=datetime(1990, 5, 5), od_end=None):
    row = {
        "District": district, "RouteNum": route, "RouteSuffix": ".",
        "Alignment": align, "BeginCounty": county,
        "EndCounty": county2 or county, "BeginPMPrefix": prefix,
        "BeginPMMeasure": b, "EndPMPrefix": prefix2 or prefix,
        "EndPMMeasure": e,
        "BeginODMeasure": od,
        "EndODMeasure": od + (e - b) if od_end is None else od_end,
        "InventoryItemStartDate": item, "RouteID": f"SHS_{route}._P",
    }
    row.update(life or _LIVE)
    row.update(attrs)
    return row


def _point(county, pm, attrs, align="Right", prefix=".", route="1", life=None):
    row = {"RouteNum": route, "RouteSuffix": ".", "Alignment": align,
           "County": county, "PMPrefix": prefix, "PMMeasure": pm}
    row.update(life or _LIVE)
    row.update(attrs)
    return row


def _write_layer(lib, nn, layer, rows, columns):
    wb = Workbook()
    ws = wb.active
    ws.title = layer[:31]
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c) for c in columns])
    path = lib / f"{nn:02d}_{layer}.xlsx"
    wb.save(path)
    return len(rows)


def _build_library(lib, *, drop_layer=None, lie_rows=None, extra=None):
    """A minimal consistent library covering every required highway layer for
    route 001 across Orange + Los Angeles. Returns the INDEX entry list.
    `extra` appends rows to one layer ({layer name: [row dicts]}) — the HF-01
    skipped-span fixtures ride the same consistent library."""
    lib.mkdir(parents=True, exist_ok=True)
    span_cols = ["OBJECTID", "District", "RouteNum", "RouteSuffix", "Alignment",
                 "BeginCounty", "EndCounty", "BeginPMPrefix", "BeginPMMeasure",
                 "BeginPMSuffix", "EndPMPrefix", "EndPMMeasure", "EndPMSuffix",
                 "BeginODMeasure", "EndODMeasure", "InventoryItemStartDate",
                 "InventoryItemEndDate", "RouteID", "LRSFromDate", "LRSToDate",
                 "LocError"]
    point_cols = ["OBJECTID", "RouteNum", "RouteSuffix", "Alignment", "County",
                  "PMPrefix", "PMMeasure", "PMSuffix", "ODMeasure",
                  "LRSFromDate", "LRSToDate"]

    ora, la = "Orange", "Los Angeles"
    # HG: ORA D 0-1, U 1-2, independent pair 2-2.5 (R) / 2-2.6 (L),
    # D 2.5-3.0, GAP 3-4 (unconstructed), D 4-5; LA D 0-1.
    hg = [
        _span(ora, 0.0, 1.0, 0.0, {"Highway_Group": "D- Divided Highway"}),
        _span(ora, 1.0, 2.0, 1.0, {"Highway_Group": "U- Undivided Highway"}),
        _span(ora, 2.0, 2.5, 2.0, {"Highway_Group": "R- Independent Alignment"}),
        _span(ora, 2.0, 2.6, 2.0, {"Highway_Group": "L- Independent Alignment"},
              align="Left"),
        _span(ora, 2.5, 3.0, 2.5, {"Highway_Group": "D- Divided Highway"}),
        _span(ora, 4.0, 5.0, 4.0, {"Highway_Group": "D- Divided Highway"}),
        _span(la, 0.0, 1.0, 5.0, {"Highway_Group": "D- Divided Highway"},
              district="District 7"),
        _span(ora, 0.0, 5.0, 0.0, {"Highway_Group": "U- Undivided Highway"},
              life=_DEAD),                       # history: never painted
        _span(la, 0.0, 1.0, 5.0, {"Highway_Group": "U- Undivided Highway"},
              district="District 7", life=_FUTURE),  # future: never painted
    ]
    whole = {
        "SHS Median": (["Median_Type", "Median_Width", "Median_Variance"],
                       {"Median_Type": "H- Paved Median", "Median_Width": 12,
                        "Median_Variance": "Z- No Variance"}),
        "Terrain Type": (["Terrain_Type"], {"Terrain_Type": "F- Level"}),
        "SHS Design Speed": (["Design_Speed"], {"Design_Speed": 65}),
        "SHS Curb Landscape": (["Curb_Landscape"],
                               {"Curb_Landscape": "7- No Curbs or Shrubs"}),
        "SHS Barrier": (["Barrier_Type"], {"Barrier_Type": "Z- No Barriers"}),
        "SHS Population": (["Population_Code"],
                           {"Population_Code": "R- Rural"}),
    }
    sides = {
        "SHS Travel Way R": (["Travel_Way_Width_R", "Total_Num_Lanes_R"],
                             {"Travel_Way_Width_R": 24, "Total_Num_Lanes_R": 2},
                             "Right"),
        "SHS Surface Type R": (["Surface_Type_R"],
                               {"Surface_Type_R": "H- AC: Base & Surface"},
                               "Right"),
        "SHS Special Feature R": (["Special_Feature_Type_R"],
                                  {"Special_Feature_Type_R":
                                   "Z- No Special Features"}, "Right"),
        "SHS O Shld Width R": (["Shld_Width_Total_Out_R",
                                "Shld_Width_Treated_Out_R"],
                               {"Shld_Width_Total_Out_R": 8,
                                "Shld_Width_Treated_Out_R": 8}, "Right"),
        "SHS I Shld Width R": (["Shld_Width_Total_In_R",
                                "Shld_Width_Treated_In_R"],
                               {"Shld_Width_Total_In_R": 0,
                                "Shld_Width_Treated_In_R": 0}, "Right"),
        "SHS Travel Way L": (["Travel_Way_Width_L", "Total_Num_Lanes_L"],
                             {"Travel_Way_Width_L": 24, "Total_Num_Lanes_L": 2},
                             "Left"),
        "SHS Surface Type L": (["Surface_Type_L"],
                               {"Surface_Type_L": "H- AC: Base & Surface"},
                               "Left"),
        "SHS Special Feature L": (["Special_Feature_Type_L"],
                                  {"Special_Feature_Type_L":
                                   "Z- No Special Features"}, "Left"),
        "SHS O Shld Width L": (["Shld_Width_Total_Out_L",
                                "Shld_Width_Treated_Out_L"],
                               {"Shld_Width_Total_Out_L": 8,
                                "Shld_Width_Treated_Out_L": 8}, "Left"),
        "SHS I Shld Width L": (["Shld_Width_Total_In_L",
                                "Shld_Width_Treated_In_L"],
                               {"Shld_Width_Total_In_L": 0,
                                "Shld_Width_Treated_In_L": 0}, "Left"),
    }

    entries, nn = [], 0

    def add(layer, rows, columns):
        nonlocal nn
        nn += 1
        if layer == drop_layer:
            return
        if extra and layer in extra:
            rows = list(rows) + list(extra[layer])
        n = _write_layer(lib, nn, layer, rows, columns)
        claimed = lie_rows.get(layer, n) if lie_rows else n
        entries.append((f"{nn:02d}_{layer}.xlsx", layer, claimed,
                        len(columns), f"path/{layer}",
                        f"https://gis.example/{layer}/FeatureServer/{nn}"))

    add("SHS Highway Group", hg, span_cols + ["Highway_Group"])
    for layer, (attr_cols, attrs) in whole.items():
        rows = [_span(ora, 0.0, 5.0, 0.0, attrs),
                _span(ora, 2.0, 2.6, 2.0, attrs, align="Left"),
                _span(la, 0.0, 1.0, 5.0, attrs, district="District 7")]
        add(layer, rows, span_cols + attr_cols)
    for layer, (attr_cols, attrs, align) in sides.items():
        rows = [_span(ora, 0.0, 5.0, 0.0, attrs, align=align),
                _span(la, 0.0, 1.0, 5.0, attrs, align=align,
                      district="District 7")]
        add(layer, rows, span_cols + attr_cols)
    # Access Control: a CROSS-COUNTY span ORA 4.5 -> LA 0.3 (odometers carry
    # the apportioning), plus plain coverage before it.
    add("SHS Access Control",
        [_span(ora, 0.0, 4.5, 0.0, {"SHS_Access_Control":
                                    "C- Conventional Highway"}),
         _span(ora, 4.5, 0.3, 4.5, {"SHS_Access_Control":
                                    "F- Freeway (full control)"},
               county2=la, od_end=5.3)],
        span_cols + ["SHS_Access_Control"])
    add("SHS Non Add Mileage",
        [_span(ora, 4.0, 4.4, 4.0, {"Non_Add_Mileage": "N - Non-Add"})],
        span_cols + ["Non_Add_Mileage"])
    add("SHS Tolls",
        [_span(ora, 0.5, 0.8, 0.5, {"Toll_Type": "Toll Roads"})],
        span_cols + ["Toll_Type"])
    add("SHS Forest HWY",
        [_span(la, 0.5, 0.8, 5.5, {"Forest_Hwy": "Yes"},
               district="District 7")],
        span_cols + ["Forest_Hwy"])
    add("SHS Inv Network Date",
        [_span(ora, 0.0, 5.0, 0.0, {"Network_Start_Date": datetime(1964, 1, 1),
                                    "SegOrderId": 100}),
         _span(la, 0.0, 1.0, 5.0, {"Network_Start_Date": datetime(1964, 1, 1),
                                   "SegOrderId": 200},
               district="District 7")],
        [c for c in span_cols if c != "InventoryItemStartDate"]
        + ["Network_Start_Date", "Network_End_Date", "SegOrderId"])
    add("City",
        [_span(ora, 0.2, 0.4, 0.2, {"City_Code": "Los Angeles"}),
         _span(ora, 0.9, 0.95, 0.9, {"City_Code": "Unincorporated Ville"}),
         _span(ora, 90.0, 91.0, 90.0, {"City_Code": "NOT SHS"})
         | {"RouteID": "ORA_X_SIDE ST_P"}],
        span_cols + ["City_Code"])
    add("Traffic Volume Segments",
        [_span(ora, 0.0, 1.0, 0.0,
               {"AADT": 1000, "AADT_YEAR": 2024, "AADT_AHEAD": 700,
                "AADT_BACK": 800}),
         _span(ora, 0.0, 1.0, 0.0,
               {"AADT": 900, "AADT_YEAR": 2022, "AADT_AHEAD": 650,
                "AADT_BACK": 750})],
        span_cols + ["AADT", "AADT_YEAR", "AADT_AHEAD", "AADT_BACK"])
    add("SHS Landmark",
        [_point(ora, 1.0, {"Landmarks_Short": "TEST LANDMARK"}),
         _point(ora, 1.0, {"Landmarks_Short": "X"})],
        point_cols + ["Landmarks_Short", "Landmarks_Long"])
    add("Equation Points",
        [_point(ora, 0.5, {"hslDescription": "EQ"})],
        point_cols + ["hslDescription"])
    add("SHS Route Break",
        [_point(ora, 4.0, {"Route_Break_Type": "Route Resume"})],
        point_cols + ["Route_Break_Type"])

    wb = Workbook()
    ws = wb.active
    ws.append(crl.INDEX_HEADER)
    for row in entries:
        ws.append(list(row))
    wb.save(lib / crl.INDEX_NAME)
    return entries


def _rows_of(path, sheet):
    wb = load_workbook(path, data_only=True)
    try:
        it = wb[sheet].iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        return header, [list(r) for r in it]
    finally:
        wb.close()


_LINE_PT = 15          # one line of the default 11pt font
_PADDING = 2           # cell inset + a margin over Excel's own font metric


def _illegible_marker_cells(path):
    """Every stored cell of the build's DISCLOSURE sheets must be READABLE at
    its column's stored width: Excel clips a label whose neighbour is occupied
    and cuts text at the column edge, so a cell must either fit its column or
    wrap in a row tall enough for every wrapped line. Numbers are exempt —
    General format rounds their display instead of clipping. Returns the
    offenders (RB-1 review 1 / RB1-R1-001)."""
    wb = load_workbook(path)
    try:
        bad = []
        sheets = [n for n in (chc.ARC_MARKER_SHEET,
                              getattr(chc, "ARC_MARKED_SHEET", ""))
                  if n and n in wb.sheetnames]
        for name in sheets:
            ws = wb[name]
            widths = {}
            for dim in ws.column_dimensions.values():
                if dim.width:
                    for i in range(dim.min, dim.max + 1):
                        widths[i] = dim.width
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None or not str(c.value).strip():
                        continue
                    where = f"{name}!{c.coordinate}"
                    width = widths.get(c.column)
                    if not width:
                        bad.append(f"{where}: column has no stored width")
                        continue
                    if isinstance(c.value, (int, float)):
                        continue
                    text = str(c.value)
                    if len(text) <= width - _PADDING:
                        continue
                    if not (c.alignment and c.alignment.wrap_text):
                        bad.append(f"{where}: {len(text)} chars in a "
                                   f"{width:g}-wide column and does not wrap")
                        continue
                    lines = len(textwrap.wrap(text,
                                              max(int(width) - _PADDING, 8)))
                    height = getattr(ws.row_dimensions.get(c.row), "height",
                                     None)
                    if height is None or height + 0.5 < lines * _LINE_PT:
                        bad.append(f"{where}: wraps to {lines} lines but the "
                                   f"row is {height} tall")
        return bad
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# unit checks
# --------------------------------------------------------------------------- #
def test_dialects_and_algebra():
    print("== dialects + LRS/PM algebra")
    check("county name -> code", crl.norm_county("Los Angeles") == "LA")
    check("county dotted code", crl.norm_county("LA.") == "LA")
    check("county code passthrough", crl.norm_county("SBD") == "SBD")
    check("district label", crl.norm_district("District 7") == "07")
    check("district number", crl.norm_district(12) == "12")
    check("alignment label", crl.norm_alignment("Right") == "R")
    check("alignment none", crl.norm_alignment(".") == "")
    check("coded label", crl.code_of("J- Unpaved Median") == "J")
    check("coded label spaced", crl.code_of("N - Non-Add") == "N")
    check("coded bare", crl.code_of("Z") == "Z")
    check("coded numeric", crl.code_of("7- No Curbs or Shrubs") == "7")
    check("dot none", crl.dot_none(".") == "")
    check("route pad", crl.norm_route("1") == "001")
    check("city name -> TASAS code", city_codes.norm_city("Los Angeles") == "LA")
    check("city code passthrough", city_codes.norm_city("SJS") == "SJS")
    check("unmapped city surfaces upper",
          city_codes.norm_city("No Such Town") == "NO SUCH TOWN")
    d = crl.to_serial(datetime(2025, 9, 8))
    check("serial of datetime", d == 45908.0)
    check("serial of iso text", crl.to_serial("2025-09-08") == 45908.0)
    check("serial of number", crl.to_serial(45908) == 45908.0)
    check("serial of dot", crl.to_serial(".") is None)
    check("as-of open current", crl.is_asof(45000.0, None, 45908.0))
    check("as-of from-boundary live", crl.is_asof(45908.0, None, 45908.0))
    check("as-of to-boundary dead", not crl.is_asof(45000.0, 45908.0, 45908.0))
    check("pm units float noise", crl.pm_units(82.71599999999999) == 8271600)
    check("pm text trims", crl.pm_text(8271600) == "82.716")
    check("pm text 5dp", crl.pm_text(18623798) == "186.23798")
    segs = crl.overlay({"A": [(0, 10, "a1", (1,)), (0, 10, "a2", (2,))],
                        "B": [(5, 15, "b", (1,))]}, cuts=(7,))
    check("overlay rank wins", segs[0][2]["A"] == "a2")
    check("overlay cut applied", any(s[0] == 7 for s in segs))
    check("overlay union edges", [s[0] for s in segs] == [0, 5, 7, 10])


def test_stream_and_index():
    print("== stream_layer + INDEX gates")
    with tempfile.TemporaryDirectory() as td:
        lib = Path(td) / "lib"
        _build_library(lib)
        inv = crl.inventory(lib)
        check("inventory sees all present", not inv["missing"] or
              set(inv["missing"]) <= (set(crl.EXPECTED_LAYERS)
                                      - set(cch.HIGHWAY_LAYERS) - {"City"}))
        check("index present", inv["index"] is not None)
        idx = crl.read_index(lib)
        check("index carries sources",
              idx["SHS Median"]["source"].startswith("https://gis.example/"))
        hg = lib / "01_SHS Highway Group.xlsx"
        rows = list(crl.stream_layer(hg, ["Highway_Group", "BeginCounty"],
                                     layer_name="SHS Highway Group"))
        check("name-keyed read", rows[0]["Highway_Group"].startswith("D-"))
        rows = list(crl.stream_layer(hg, ["Missing_Col"], optional=("Missing_Col",)))
        check("optional column reads None", rows[0]["Missing_Col"] is None)
        try:
            list(crl.stream_layer(hg, ["Nope"]))
            check("missing wanted column refuses", False)
        except ValueError as e:
            check("missing wanted column refuses", "Nope" in str(e))
        try:
            list(crl.stream_layer(hg, ["Highway_Group"], expected_rows=99))
            check("truncated export refuses", False)
        except ValueError as e:
            check("truncated export refuses", "truncated" in str(e))
        n = len(list(crl.stream_layer(hg, ["Highway_Group"])))
        ok = list(crl.stream_layer(hg, ["Highway_Group"], expected_rows=n - 2))
        check("healthy over-count race passes", len(ok) == n)


def test_consolidator_end_to_end():
    print("== consolidator end-to-end (synthetic library)")
    with tempfile.TemporaryDirectory() as td:
        lib = Path(td) / "lib"
        _build_library(lib)
        out = Path(td) / "built.xlsx"
        res = cch.consolidate(events=Events(), asof=ASOF, lib_root=lib,
                              out_path=out)
        check("build ok", res.status == "ok")
        check("build complete", res.completion == "complete")
        check("build names output", res.output_path == str(out))
        header, rows = _rows_of(out, chc.ARC_SHEET)
        check("74-column header", header == chc.HEADER)
        col = {n: i for i, n in enumerate(chc.HEADER)}

        def rows_where(**kw):
            keep = []
            for r in rows:
                if all(str(r[col[k]] or "") == str(v) for k, v in kw.items()):
                    keep.append(r)
            return keep

        ora = rows_where(THY_COUNTY_CODE="ORA")
        la = rows_where(THY_COUNTY_CODE="LA")
        check("both counties built", bool(ora) and bool(la))
        check("district painted per county",
              ora[0][col["THY_DISTRICT_CODE"]] == "12"
              and la[0][col["THY_DISTRICT_CODE"]] == "07")
        check("an HG coverage gap yields NO fabricated row (TSN's X rows "
              "stay one-sided)",
              not any(3.0 <= r[col["THY_BEGIN_PM_AMT"]] < 4.0 for r in ora))
        r_rows = rows_where(THY_PM_SUFFIX_CODE="R")
        l_rows = rows_where(THY_PM_SUFFIX_CODE="L")
        check("independent pair rows exist", r_rows and l_rows)
        check("R row nulls the LT block",
              r_rows[0][col["THY_LT_SURF_TYPE_CODE"]] is None
              and r_rows[0][col["THY_RT_SURF_TYPE_CODE"]] == "H")
        check("L row nulls the RT block",
              l_rows[0][col["THY_RT_SURF_TYPE_CODE"]] is None
              and l_rows[0][col["THY_LT_SURF_TYPE_CODE"]] == "H")
        check("L row spans its own alignment PMs",
              l_rows[0][col["THY_BEGIN_PM_AMT"]] == 2.0
              and l_rows[-1][col["THY_END_PM_AMT"]] == 2.6)
        first = ora[0]
        check("profile anchor: P + AADT_AHEAD at the span begin",
              first[col["THY_PROFILE_CODE"]] == "P"
              and first[col["THY_ADT_AMT"]] == 700
              and first[col["THY_CHANGE_PER_MILE_AMT"]] == 100.0)
        cut_rows = [r for r in ora if r[col["THY_BEGIN_PM_AMT"]] in (0.2, 0.4)]
        check("city boundaries cut rows", len(cut_rows) == 2)
        check("equate point cuts + flags",
              any(r[col["THY_BEGIN_PM_AMT"]] == 0.5
                  and r[col["THY_EQUATE_CODE"]] == "E" for r in ora))
        lmk = [r for r in ora if r[col["THY_BEGIN_PM_AMT"]] == 1.0]
        check("landmark attaches (longest text)",
              lmk and lmk[0][col["THY_LANDMARK_SHORT_DESC"]] == "TEST LANDMARK")
        toll = [r for r in ora if r[col["THY_BEGIN_PM_AMT"]] == 0.5]
        check("toll span -> code 1",
              toll and toll[0][col["THY_TOLL_FOREST_CODE"]] == 1)
        forest = [r for r in la if r[col["THY_BEGIN_PM_AMT"]] == 0.5]
        check("forest span -> code 2",
              forest and forest[0][col["THY_TOLL_FOREST_CODE"]] == 2)
        check("non-add span -> N; default A",
              any(r[col["THY_NON_ADD_CODE"]] == "N" for r in ora)
              and la[0][col["THY_NON_ADD_CODE"]] == "A")
        check("cross-county access reaches LA",
              la[0][col["THY_HIGHWAY_ACCESS_CODE"]] == "F")
        check("cross-county access tail stays in ORA",
              any(r[col["THY_BEGIN_PM_AMT"]] == 4.5
                  and r[col["THY_HIGHWAY_ACCESS_CODE"]] == "F" for r in ora))
        check("no-source columns stay empty",
              all(r[col["THY_MAINT_SVC_LVL_CODE"]] is None for r in rows))
        in_city = [r for r in ora
                   if 0.2 <= r[col["THY_BEGIN_PM_AMT"]] < 0.4]
        check("city names normalize to TASAS codes",
              in_city and all(r[col["THY_CITY_CODE"]] == "LA"
                              for r in in_city))
        unmapped = [r for r in ora
                    if 0.9 <= r[col["THY_BEGIN_PM_AMT"]] < 0.95]
        check("an unmapped city name passes through visibly (upper-cased)",
              unmapped and all(r[col["THY_CITY_CODE"]]
                               == "UNINCORPORATED VILLE" for r in unmapped))
        check("no city outside the spans",
              all(r[col["THY_CITY_CODE"]] is None for r in ora
                  if r[col["THY_BEGIN_PM_AMT"]] >= 1.0))
        check("extract date = the as-of date",
              str(rows[0][col["THY_EXTRACT_DATE"]]).startswith(ASOF))
        offs = [r[col["THY_BEGIN_OFFSET_AMT"]] for r in ora]
        check("offsets are monotone within the route",
              all(a <= b for a, b in zip(offs, offs[1:])))
        check("first-county offsets ARE the begin PMs (PM-continued; the "
              "parallel L roadbed reads but never advances the line)",
              all(r[col["THY_BEGIN_OFFSET_AMT"]]
                  == r[col["THY_BEGIN_PM_AMT"]] for r in ora
                  if r[col["THY_PM_SUFFIX_CODE"]] != "L"))
        check("the county line continues the cumulative (LA starts at ORA's "
              "corridor end)",
              la[0][col["THY_BEGIN_OFFSET_AMT"]]
              == 5.0 + la[0][col["THY_BEGIN_PM_AMT"]])
        check("BEG marks the route start",
              rows[0][col["THY_BREAK_DESC"]] == "BEG")
        check("route-break point -> U-BR (resume)",
              any(r[col["THY_BREAK_DESC"]] == "U-BR" for r in ora))

        pheader, prows = _rows_of(out, "Provenance")
        check("provenance covers all 74 columns",
              len(prows) == len(chc.HEADER)
              and [r[0] for r in prows] == chc.HEADER)
        check("provenance carries FeatureServer sources",
              any("https://gis.example/" in str(r[5] or "") for r in prows))
        mheader, mrows = _rows_of(out, chc.ARC_MARKER_SHEET)
        marker = {str(mheader[0]): mheader[1]}
        for r in mrows:
            marker[str(r[0])] = r[1]
        check("build marker carries the as-of", marker.get("As-of date") == ASOF)

        # Header tinting (owner ask 2026-07-27): un-sourceable columns carry
        # the Provenance grey + a hover note on the DATA sheet's header; every
        # sourced/synthesized header stays plain. Presentation only.
        swb = load_workbook(out)
        try:
            hdr = {c.value: c for c in swb[chc.ARC_SHEET][1]
                   if c.value is not None}

            def _fill_of(name):
                f = hdr[name].fill
                return (str(f.start_color.rgb)
                        if f is not None and f.patternType else "")

            check("no-source header tinted grey + noted",
                  _fill_of("THY_MAINT_SVC_LVL_CODE").endswith("D9D9D9")
                  and hdr["THY_MAINT_SVC_LVL_CODE"].comment is not None)
            check("TSN-internal header tinted light grey",
                  _fill_of("THY_ID").endswith("EDEDED"))
            check("sourced header stays plain",
                  _fill_of("THY_COUNTY_CODE") == "")
            check("synthesized header stays plain",
                  _fill_of("THY_BEGIN_OFFSET_AMT") == "")
        finally:
            swb.close()

        # refusals
        lib2 = Path(td) / "lib2"
        _build_library(lib2, drop_layer="SHS Median")
        res2 = cch.consolidate(events=Events(), asof=ASOF, lib_root=lib2,
                               out_path=Path(td) / "b2.xlsx")
        check("missing layer refuses by name",
              res2.status == "error" and "SHS Median" in res2.message)
        lib3 = Path(td) / "lib3"
        _build_library(lib3, lie_rows={"SHS Barrier": 99})
        res3 = cch.consolidate(events=Events(), asof=ASOF, lib_root=lib3,
                               out_path=Path(td) / "b3.xlsx")
        check("truncated layer refuses via the INDEX gate",
              res3.status == "error" and "truncated" in res3.message)
        res4 = cch.consolidate(events=Events(), asof="nonsense", lib_root=lib,
                               out_path=Path(td) / "b4.xlsx")
        check("bad as-of refuses", res4.status == "error"
              and "as-of" in res4.message)


def _tsn_raw(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = chc.TSN_RAW_SHEET
    ws.append(list(chc.HEADER))
    for r in rows:
        ws.append(r)
    wb.save(path)


def _thy_row(county="ORA", route="001", begin=0.0, end=1.0, maint=None,
             hg="D"):
    row = [None] * len(chc.HEADER)
    col = {n: i for i, n in enumerate(chc.HEADER)}
    row[col["THY_DISTRICT_CODE"]] = "12"
    row[col["THY_COUNTY_CODE"]] = county
    row[col["THY_ROUTE_NAME"]] = route
    row[col["THY_BEGIN_PM_AMT"]] = begin
    row[col["THY_END_PM_AMT"]] = end
    row[col["THY_LENGTH_MILES_AMT"]] = round(end - begin, 5)
    row[col["THY_HIGHWAY_GROUP_CODE"]] = hg
    row[col["THY_MAINT_SVC_LVL_CODE"]] = maint
    row[col["THY_EXTRACT_DATE"]] = datetime(2025, 9, 8)
    return row


def test_normalizer_and_comparator():
    print("== TSN normalizer + comparator role gates + context columns")
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        lib = td / "lib"
        _build_library(lib)
        built = td / "built.xlsx"
        res = cch.consolidate(events=Events(), asof=ASOF, lib_root=lib,
                              out_path=built)
        check("arc side built", res.status == "ok")

        raw_dir = td / "raw"
        raw_dir.mkdir()
        # The TSN side mirrors the built ORA base rows loosely: same first
        # two rows, a MAINT value everywhere (context — never counted), and
        # one REAL difference (HG on the second row).
        header, arows = _rows_of(built, chc.ARC_SHEET)
        col = {n: i for i, n in enumerate(chc.HEADER)}
        tsn_rows = []
        for i, r in enumerate(arows):
            rr = list(r)
            rr[col["THY_MAINT_SVC_LVL_CODE"]] = 2          # context-only delta
            if i == 1:
                rr[col["THY_HIGHWAY_GROUP_CODE"]] = "Q"    # ONE real diff
            tsn_rows.append(rr)
        _tsn_raw(raw_dir / "CA HIGHWAYS test.xlsx", tsn_rows)

        norm = td / "normalized.xlsx"
        nres = tlc.build_into_highway(raw_dir, norm, events=Events())
        check("normalizer ok", nres.status == "ok")
        wb = load_workbook(norm)
        try:
            import compare_tsn_common as ctc
            check("normalized sheet present",
                  chc.NORMALIZED_SHEET in wb.sheetnames)
            check("normalization marker v1",
                  ctc.normalization_marker_version(wb)
                  == cht.NORMALIZATION_VERSION)
        finally:
            wb.close()

        bad = td / "bad_raw"
        bad.mkdir()
        wbx = Workbook()
        wsx = wbx.active
        wsx.title = chc.TSN_RAW_SHEET
        wsx.append(list(chc.HEADER[:-1]) + ["WRONG"])
        wsx.append([None] * len(chc.HEADER))
        wbx.save(bad / "CA HIGHWAYS bad.xlsx")
        nbad = tlc.build_into_highway(bad, td / "n2.xlsx", events=Events())
        check("normalizer refuses a drifted header", nbad.status == "error")

        # Role gates.
        try:
            cht._load_arc(norm)
            check("ARC side refuses an unmarked/TSN workbook", False)
        except ValueError as e:
            check("ARC side refuses an unmarked/TSN workbook",
                  chc.ARC_MARKER_SHEET in str(e))
        try:
            cht._load_tsn(built)
            check("TSN side refuses the ArcGIS build", False)
        except ValueError as e:
            check("TSN side refuses the ArcGIS build",
                  chc.ARC_MARKER_SHEET in str(e))
        check("TSN side loads the normalized library",
              len(cht._load_tsn(norm)) == len(tsn_rows))
        check("TSN side loads the raw extract",
              len(cht._load_tsn(raw_dir / "CA HIGHWAYS test.xlsx"))
              == len(tsn_rows))

        # The full comparison, both flavors, on the shipped path.
        out = td / "cmp.xlsx"
        cres = cht.compare(built, norm, out, events=Events(), mode="both")
        check("comparison ok", cres.status == "ok")
        values_twin = out.with_name(out.stem + " (values)" + out.suffix)
        check("both flavors written", out.is_file() and values_twin.is_file())
        oc = cres.comparison_outcome
        counts = getattr(oc, "counts", None)
        check("typed outcome carries counts",
              counts is not None and counts.known)
        diffs = getattr(counts, "differing_cells", None)
        check("exactly the ONE real difference is counted (context never "
              f"counts) — got {diffs}", diffs == 1)

        # M2-E item 10 (v0.32.0): the 24 context (shown-only) column HEADERS
        # are tinted grey with a hover note on the Comparison sheet, visibly
        # distinct from the counted columns — presentation only (the count
        # assertions above are the same run).
        vwb = load_workbook(values_twin)
        try:
            cws = vwb["Comparison"]
            hdr = {c.value: c for c in cws[1] if c.value is not None}
            ctx_field = chc.CONTEXT_COLUMNS[0]
            cmp_field = next(f for f in cht.SHARED_HEADER
                             if f not in chc.CONTEXT_COLUMNS and f != cht.KEY)
            ctx_cell, cmp_cell = hdr.get(ctx_field), hdr.get(cmp_field)
            check("context header present on the Comparison sheet",
                  ctx_cell is not None and cmp_cell is not None)
            check("context header tinted with the schema's fill",
                  ctx_cell is not None
                  and str(ctx_cell.fill.start_color.rgb).endswith("808080"))
            check("compared header keeps the standard band",
                  cmp_cell is not None
                  and str(cmp_cell.fill.start_color.rgb).endswith("1F3864"))
            check("context header carries the hover note",
                  ctx_cell is not None and ctx_cell.comment is not None
                  and "never" in ctx_cell.comment.text)
        finally:
            vwb.close()


def test_skipped_span_source_truth():
    """HF-01 / PCOA-FINAL-010 red->green: an as-of span with one unreadable
    postmile endpoint must be RECORDED (sidecar + marker + PARTIAL), its
    known-endpoint anchor cells MARKED with the reserved unavailable token
    (never guessed, never blank), and the marked anchors must never count as
    differences in a real mode="both" comparison — while a genuinely
    differing, correctly placed control cell still counts."""
    print("== skipped-source spans: recorded, PARTIAL, marked, never counted"
          " (HF-01)")
    # getattr + literal so the check DEGRADES TO RED (not a crash) on the
    # pre-fix code, which has no token at all — the red->green property.
    tok = getattr(cch, "UNAVAILABLE_TOKEN", "(unavailable: source span skipped)")
    ora = "Orange"
    osr = {"Shld_Width_Total_Out_R": 8, "Shld_Width_Treated_Out_R": 8}
    # Three OSR spans the postmile contract cannot place (LocError=NO ERROR,
    # usable odometers — exactly the witnessed raw shape):
    #  - begin known @0.7, Total differs from the painted 8 -> ONE marked cell
    #    (Treated equals the painted value -> corroborated, NOT marked);
    #  - begin known @1.5, both values equal the painted ones -> recorded,
    #    nothing marked;
    #  - end known @2.2 (begin unreadable) inside the R-window -> marks the
    #    R row's Total (kind eligibility: base/R rows carry the RT block);
    #  - a SECOND end-known @2.2 span carrying a different Total -> the same
    #    cell now stands in front of TWO unplaceable values (the real 036/TEH
    #    shape), one of which is the value TSN shows.
    skip_mark = _span(ora, 0.7, None, 0.7,
                      {"Shld_Width_Total_Out_R": 4, "Shld_Width_Treated_Out_R": 8,
                       "LocError": "NO ERROR"}, od_end=0.9)
    skip_match = _span(ora, 1.5, None, 1.5,
                       dict(osr, LocError="NO ERROR"), od_end=1.7)
    skip_end = _span(ora, None, 2.2, None,
                     {"Shld_Width_Total_Out_R": 4, "Shld_Width_Treated_Out_R": 8,
                      "LocError": "NO ERROR"}, od_end=2.4)
    skip_end2 = _span(ora, None, 2.2, None,
                      {"Shld_Width_Total_Out_R": 5, "Shld_Width_Treated_Out_R": 8,
                       "LocError": "NO ERROR"}, od_end=2.6)
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        lib = td / "lib"
        _build_library(lib, extra={"SHS O Shld Width R":
                                   [skip_mark, skip_match, skip_end,
                                    skip_end2]})
        out = td / "built.xlsx"
        res = cch.consolidate(events=Events(), asof=ASOF, lib_root=lib,
                              out_path=out)
        check("build still ok", res.status == "ok")
        check("build reports PARTIAL (a skipped span is incomplete coverage)",
              res.completion == "partial")
        check("skipped_inputs counts the four spans", res.skipped_inputs == 4)
        check("result message names the marked anchors",
              "anchor cell(s)" in (res.message or ""))

        # (c) marker sheet + sidecar carry the record.
        _mh, mrows = _rows_of(out, chc.ARC_MARKER_SHEET)
        marker = {}
        for r in mrows:
            if r and r[0] is not None:
                marker.setdefault(str(r[0]), []).append(r[1])
        check("marker: skipped source spans = 4",
              marker.get("Skipped source spans") == [4])
        check("marker: marked anchor cells = 2",
              marker.get("Marked anchor cells") == [2])
        check("marker: names the unavailable token",
              marker.get("Unavailable marker") == [tok])
        check("marker: states the skip reason",
              "postmile" in str(marker.get("Skipped source reason", [""])[0]))
        # The itemized table: one readable column per field, so the 102-span
        # statewide record is scannable instead of 102 long sentences.
        want_head = [n for n, _w in getattr(cch, "_SKIP_TABLE_COLUMNS", ())]
        head_at = [i for i, r in enumerate(mrows)
                   if r and str(r[0]) == "Source layer"]
        trows = ([r for r in mrows[head_at[0] + 1:] if r and r[0]]
                 if head_at else [])
        check("marker: the skip table names every field",
              bool(head_at) and [str(c) for c in mrows[head_at[0]]
                                 if c is not None] == want_head)
        check("marker: one itemized table row per skipped span",
              len(trows) == 4
              and all(r[0] == "SHS O Shld Width R" and r[2] == "001"
                      and r[4] == "ORA" and r[8] == "NO ERROR"
                      and "Shld_Width_Total_Out_R=" in str(r[1])
                      for r in trows))
        check("marker: the table's known PM and marked counts are exact",
              sorted(str(r[7]) for r in trows) == ["0.7", "1.5", "2.2", "2.2"]
              and sorted(r[13] for r in trows) == [0, 1, 1, 1])
        check("marker: points at the itemized marked-anchor sheet",
              marker.get("Marked anchor detail")
              == [getattr(chc, "ARC_MARKED_SHEET", "")])
        check("marker: every disclosure cell is legible at its stored width",
              _illegible_marker_cells(out) == [])
        sidecar = json.loads(
            (out.parent / (out.name + ".outcome.json")).read_text("utf-8"))
        rec = (sidecar.get("clean_road_build") or {}).get(
            "skipped_source_spans") or {}
        check("sidecar: count/marked/reason recorded",
              rec.get("count") == 4 and rec.get("marked_anchor_cells") == 2
              and "postmile" in str(rec.get("reason")))
        spans = rec.get("spans") or []
        check("sidecar: spans carry layer/route/county/anchor/measures",
              len(spans) == 4
              and all(s.get("layer") == "SHS O Shld Width R"
                      and s.get("route") == "001" and s.get("county") == "ORA"
                      and s.get("loc_error") == "NO ERROR"
                      and s.get("station_pm") for s in spans))
        check("sidecar: per-span marked-cell counts are exact",
              sorted(s.get("marked_cells") for s in spans) == [0, 1, 1, 1])

        # (d) the anchors themselves: marked where information was omitted,
        # untouched where the placeable coverage corroborates the span.
        _h, rows = _rows_of(out, chc.ARC_SHEET)
        col = {n: i for i, n in enumerate(chc.HEADER)}
        tot, trt = col["THY_RT_O_SHD_TOT_WIDTH_AMT"], col["THY_RT_O_SHD_TRT_WIDTH_AMT"]
        token_cells = [(i, j) for i, r in enumerate(rows)
                       for j, v in enumerate(r) if v == tok]
        check("exactly the two omitted anchors are marked",
              len(token_cells) == 2
              and all(j == tot for _i, j in token_cells))
        anchor_a = [r for r in rows
                    if r[col["THY_BEGIN_PM_AMT"]] is not None
                    and r[col["THY_BEGIN_PM_AMT"]] <= 0.7
                    and (r[col["THY_END_PM_AMT"]] or 0) > 0.7
                    and r[col["THY_PM_SUFFIX_CODE"]] is None]
        check("begin anchor: containing base row Total is the token, the "
              "corroborated Treated keeps its value",
              anchor_a and anchor_a[0][tot] == tok and anchor_a[0][trt] == 8)
        anchor_b = [r for r in rows if r[col["THY_PM_SUFFIX_CODE"]] == "R"
                    and r[tot] == tok]
        check("end anchor: the R row inside the window is the marked one",
              len(anchor_b) == 1)
        check("the matches-raw span marked nothing at 1.5",
              all(r[tot] != tok for r in rows
                  if r[col["THY_BEGIN_PM_AMT"]] is not None
                  and 1.0 <= r[col["THY_BEGIN_PM_AMT"]] < 2.0))

        # The itemized per-cell record: WHERE every marker sits and WHAT it
        # withholds. The comparison joins it to TSN to name the anchors whose
        # withheld value TSN disagrees with (RB-1 review 2 / RB1-R2-001), so a
        # missing or wrong row here is a missing user-facing source fact.
        marked_rows = []
        _wbm = load_workbook(out, read_only=True, data_only=True)
        try:
            _mname = getattr(chc, "ARC_MARKED_SHEET", "")
            if _mname and _mname in _wbm.sheetnames:
                marked_rows = [list(r) for r in
                               _wbm[_mname].iter_rows(values_only=True)]
        finally:
            _wbm.close()
        check("marked-anchor sheet: header names every field",
              bool(marked_rows)
              and [str(c) for c in marked_rows[0] if c is not None]
              == [n for n, _w in getattr(cch, "_MARKED_TABLE_COLUMNS", ())])
        body = [r for r in marked_rows[1:] if r and r[0]]
        check("marked-anchor sheet: a row per (marked cell, withholding span)",
              len(body) == 3
              and all(r[0] == "001" and r[1] == "ORA"
                      and r[5] == "THY_RT_O_SHD_TOT_WIDTH_AMT"
                      and r[7] == "SHS O Shld Width R" for r in body))
        check("marked-anchor sheet: the withheld values and known PMs are "
              "exact (the co-anchored cell keeps BOTH)",
              sorted((r[6], str(r[8])) for r in body)
              == [(4, "0.7"), (4, "2.2"), (5, "2.2")])
        check("marked-anchor sheet: the begin postmiles ARE the anchor rows'",
              bool(anchor_a) and bool(anchor_b)
              and sorted({str(r[3]) for r in body})
              == sorted({str(r[col["THY_BEGIN_PM_AMT"]])
                         for r in (anchor_a[0], anchor_b[0])}))
        check("marked-anchor sheet: both roadbeds are named",
              sorted({"" if r[4] is None else str(r[4])
                      for r in body}) == ["", "R"])

        # The comparison: token vs the SAME value the span carried (the
        # false-positive class, TSN=4 at the begin anchor) and token vs a
        # DIFFERENT value (the misrepresented class, TSN=5 at the end anchor)
        # are both explicit N — never counted; the correctly placed control
        # difference still counts. The TSN edits key on the ANCHOR POSITION,
        # not on the token, so on the pre-fix code (blank anchors) the same
        # fixture reproduces the counted false positives — the red.
        header, arows = _rows_of(out, chc.ARC_SHEET)
        tsn_rows = []
        for i, r in enumerate(arows):
            rr = list(r)
            begin = rr[col["THY_BEGIN_PM_AMT"]]
            end = rr[col["THY_END_PM_AMT"]] or 0
            suffix = rr[col["THY_PM_SUFFIX_CODE"]]
            if begin is not None and rr[col["THY_COUNTY_CODE"]] == "ORA":
                if suffix is None and begin <= 0.7 < end:
                    rr[tot] = 4                      # == the skipped span's raw
                elif suffix == "R" and begin < 2.2 <= end:
                    rr[tot] = 5                      # != the skipped span's raw
            if i == 1:
                rr[col["THY_HIGHWAY_GROUP_CODE"]] = "Q"    # the control diff
            tsn_rows.append(rr)
        raw_dir = td / "raw"
        raw_dir.mkdir()
        _tsn_raw(raw_dir / "CA HIGHWAYS test.xlsx", tsn_rows)
        cmp_out = td / "cmp.xlsx"
        cres = cht.compare(out, raw_dir / "CA HIGHWAYS test.xlsx", cmp_out,
                           events=Events(), mode="both")
        check("comparison ok", cres.status == "ok")
        counts = getattr(cres.comparison_outcome, "counts", None)
        diffs = getattr(counts, "differing_cells", None)
        check("unavailable anchors are never counted — only the control "
              f"difference is (got {diffs})", diffs == 1)

        values_twin = cmp_out.with_name(cmp_out.stem + " (values)"
                                        + cmp_out.suffix)
        vwb = load_workbook(values_twin, data_only=True)
        try:
            cmp_texts = [str(c.value) for row in vwb["Comparison"].iter_rows()
                         for c in row if c.value is not None]
            shown = [t for t in cmp_texts if tok in t]
            check("both token anchors display the token on the Comparison "
                  "sheet", len(shown) == 2)
            check("no token anchor is displayed as a difference",
                  not any("≠" in t for t in shown))
            summary = [str(c.value) for row in vwb["Summary"].iter_rows()
                       for c in row if c.value is not None]
            check("Summary disclosure: skipped-span count + marked anchors",
                  any("SOURCE COVERAGE" in t and "4 source span(s)" in t
                      and "2 anchor cell(s)" in t for t in summary))
            check("Summary carries the producer's PARTIAL input note",
                  any("producer outcome is 'partial'" in t for t in summary))
            notes = [str(c.value) for row in vwb["Notes"].iter_rows()
                     for c in row if c.value is not None]
            check("Notes disclosure: reason + non-asserting rule",
                  any("SOURCE COVERAGE" in t and "postmile" in t
                      for t in notes)
                  and any("NON-ASSERTING" in t for t in notes))
            # RB1-R2-001: the marker must not HIDE the source fact it stands
            # in front of. The end anchor withholds 4 where TSN says 5 — that
            # exact identity must be itemized in BOTH sheets, and the begin
            # anchor (whose withheld 4 equals TSN's 4) must NOT be.
            conflict = ("001", "ORA", "THY_RT_O_SHD_TOT_WIDTH_AMT",
                        "ArcGIS source 4 @ 2.2", "TSN 5")
            check("Summary itemizes the raw-source disagreement by identity",
                  any(all(p in t for p in conflict) for t in summary))
            check("Notes itemizes the same disagreement by identity",
                  any(all(p in t for p in conflict) for t in notes))
            check("both sheets classify the marked anchors 1 same / 1 differs",
                  all(any("1 withhold only the value TSN already shows" in t
                          and "1 withhold at least one value TSN does not" in t
                          for t in texts) for texts in (summary, notes)))
            check("the corroborated anchor is never itemized",
                  sum(1 for t in notes if ": ArcGIS source " in t) == 1)
            check("a co-anchored marker names EVERY value it stands in front "
                  "of, not just the nearest",
                  any("ArcGIS source 4 @ 2.2, 5 @ 2.2" in t
                      for t in summary + notes))
        finally:
            vwb.close()

        # An older build without the marker record still compares (the plain
        # schema), and a skip-free build still reports COMPLETE.
        clean_lib = td / "clean_lib"
        _build_library(clean_lib)
        clean_out = td / "clean_built.xlsx"
        res2 = cch.consolidate(events=Events(), asof=ASOF, lib_root=clean_lib,
                               out_path=clean_out)
        check("a skip-free build stays COMPLETE",
              res2.status == "ok" and res2.completion == "complete")
        check("a skip-free marker records zeroes",
              getattr(cht, "_build_skip_facts",
                      lambda _p: None)(clean_out) == (0, 0, {}))
        check("a skip-free marker is legible too",
              _illegible_marker_cells(clean_out) == [])


def main():
    test_dialects_and_algebra()
    test_stream_and_index()
    test_consolidator_end_to_end()
    test_normalizer_and_comparator()
    test_skipped_span_source_truth()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL CLEAN-ROAD CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
