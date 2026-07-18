"""Golden check for CMP-AUD-046 — position-shifted exports get corrected field labels.

Ramp Detail and Intersection Detail exports have header labels shifted relative to
their VALUE positions, so the cross-environment comparison reported a real Ramp
Description change under "R/U" and an Intersection INT Type change under
"INT Eff-Date" — the diff fired correctly (values compare by position) but pointed
the user at the wrong business attribute.

After:
  * Ramp Detail (Excel + PDF) applies a position-authoritative `force_header` — one
    corrected label per VALUE position (like Highway Log's corrected labels);
  * Intersection Detail's current site edition realigned its labels over their
    values, and `_id_canonical_header` maps the legacy export to it: the Excel side
    got this via CMP-AUD-032/048, and the PDF side now pins the same canonicalizer.

Values are compared BY POSITION, so no difference COUNT changes — only the DISPLAYED
field name is corrected. Censused against the real 7.9 statewide exports (RD/ID,
Excel + converted PDF).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_env_field_labels.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
from compare_core import _DIFF_MARK
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []


def check(name, ok, extra=""):
    print(f"  [{'OK ' if ok else 'FAIL'}] {name}" + (f"  {extra}" if extra else ""))
    if not ok:
        _fail.append(name)


# The censused corrected Ramp Detail label per VALUE position (file, no Route).
_RD_CORRECT = ["Location", "PR", "PM", "Date of Record", "PM Suffix", "HG",
               "Area 4", "City Code", "R/U", "Description", "(unused)"]


def test_config_wiring():
    print("Config wiring (force_header / canonicalizer per family):")
    check("Ramp Detail force_header = the corrected labels",
          list(compare_env.RAMP_DETAIL.force_header) == _RD_CORRECT)
    check("Ramp Detail (PDF) force_header = corrected + the two print columns",
          list(compare_env.RAMP_DETAIL_PDF.force_header)
          == _RD_CORRECT + ["On/Off", "Ramp Type"])
    check("Intersection Detail (PDF) pins the ID canonicalizer",
          compare_env.INTERSECTION_DETAIL_PDF.header_canonicalizer
          is compare_env._id_canonical_header)


def test_id_realignment():
    print("Intersection Detail label realignment (legacy -> current, over values):")
    import compare_intersection_detail_tsn as idt
    canon = compare_env._id_canonical_header
    legacy = list(idt._TSMIS_HEADER_LEGACY[1:])
    current = canon(legacy)
    # The finding's example: legacy pos 8/9 label INT Type / INT Eff-Date over the
    # eff-date / INT-Type values (swapped); current realigns them.
    check("legacy INT Type/INT Eff-Date labels are the swapped pair",
          legacy[8] == "INT Type" and legacy[9] == "INT Eff-Date")
    check("current label at the INT-Type VALUE position (9) is INT Type",
          current[9] == "INT Type")
    check("current label at the eff-date VALUE position (8) is the eff-date",
          "Eff" in current[8] and current[8] != "INT Type")


def _write(folder, sheet, route, header, rows):
    folder.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    wb.save(folder / f"x_route_{route}.xlsx")
    wb.close()


def _comparison(out):
    wb = load_workbook(out, read_only=True, data_only=True)
    rows = list(wb["Comparison"].iter_rows(values_only=True))
    wb.close()
    header = [("" if c is None else str(c)) for c in rows[0]]
    body = [[("" if c is None else str(c)) for c in r] for r in rows[1:]]
    return header, body


def _diff_columns(header, body):
    """The set of Comparison column headers whose cell carries the diff mark."""
    cols = set()
    for r in body:
        for i, v in enumerate(r):
            if _DIFF_MARK in v and i < len(header):
                cols.add(header[i])
    return cols


def _rd_row(pm, desc, ru="U"):
    row = [""] * 11
    row[0] = "12-ORA-001"          # Location
    row[1] = "R"                   # PR (unnamed on disk)
    row[2] = pm                    # PM
    row[8] = ru                    # R/U VALUE (label "City Code" on disk)
    row[9] = desc                  # Description VALUE (label "R/U" on disk)
    return row


def test_rd_description_diff_under_description():
    print("RD end-to-end: a Description change shows under 'Description', not 'R/U':")
    import compare_ramp_detail_tsn as rd
    root = Path(tempfile.mkdtemp(prefix="rd046_"))
    a = root / "2026-06-19 ssor-prod" / "ramp_detail"
    b = root / "2026-06-19 ars-prod" / "ramp_detail"
    hdr = list(rd._TSMIS_HEADER[1:])          # the real shifted export header
    _write(a, "TSAR - Ramp Detail", "001", hdr, [_rd_row("1.000", "DESC A")])
    _write(b, "TSAR - Ramp Detail", "001", hdr, [_rd_row("1.000", "DESC B")])  # Description differs
    out = root / "cmp.xlsx"
    res = compare_env.RAMP_DETAIL.compare_folders(
        str(a.parent), str(b.parent), str(out), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok", res.message)
    header, body = _comparison(out)
    check("the corrected labels are the display header (Description present, at value pos)",
          "Description" in header and "R/U" in header)
    diffcols = _diff_columns(header, body)
    check("the diff is shown under 'Description'", "Description" in diffcols, str(diffcols))
    check("the diff is NOT shown under 'R/U'", "R/U" not in diffcols, str(diffcols))


def _id_row(pm, intt, desc="MAIN ST"):
    row = [""] * 35
    row[0] = "R"                   # P / PP (postmile prefix)
    row[1] = pm                    # Post Mile
    row[3] = "12 ORA 001"          # Location
    row[4] = "73-10-19"            # Date of Record
    row[9] = intt                  # INT Type VALUE (legacy label "INT Eff-Date")
    row[20] = desc                 # Description
    return row


def test_id_type_diff_under_int_type():
    print("ID end-to-end: an INT Type change shows under 'INT Type', not 'INT Eff-Date':")
    import compare_intersection_detail_tsn as idt
    root = Path(tempfile.mkdtemp(prefix="id046_"))
    a = root / "2026-06-19 ssor-prod" / "intersection_detail"
    b = root / "2026-06-19 ars-prod" / "intersection_detail"
    hdr = list(idt._TSMIS_HEADER_LEGACY[1:])      # the real LEGACY export header
    _write(a, "Intersection Detail", "001", hdr, [_id_row("1.000", "T")])
    _write(b, "Intersection Detail", "001", hdr, [_id_row("1.000", "F")])  # INT Type T->F
    out = root / "cmp.xlsx"
    res = compare_env.INTERSECTION_DETAIL.compare_folders(
        str(a.parent), str(b.parent), str(out), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok", res.message)
    header, body = _comparison(out)
    check("the display header uses the current realigned labels (INT Type present)",
          "INT Type" in header)
    diffcols = _diff_columns(header, body)
    check("the diff is shown under 'INT Type'", "INT Type" in diffcols, str(diffcols))
    check("the diff is NOT shown under 'INT Eff-Date'",
          "INT Eff-Date" not in diffcols, str(diffcols))


def main():
    test_config_wiring()
    test_id_realignment()
    test_rd_description_diff_under_description()
    test_id_type_diff_under_int_type()
    if _fail:
        print(f"\nFAILED: {len(_fail)}")
        raise SystemExit(1)
    print("\nALL COMPARE-ENV-FIELD-LABELS (CMP-AUD-046) CHECKS PASSED")


if __name__ == "__main__":
    main()
