"""Golden check for the Intersection consolidators (v0.17.0).

Covers consolidate_intersection_detail (a thin wrapper over the shared
consolidate_xlsx core): the config that can't drift (subdir / sheet name /
output filename) and an end-to-end 2-route consolidation read back with openpyxl
(leading Route column, header locked, rows combined). The free-text Description
column's formula-injection guard lives in the shared consolidate_xlsx core and is
locked separately by check_compare_injection.

Also covers consolidate_intersection_summary (v0.17.0): the block-walk category
summer — a 2-route consolidation read back, verifying the per-route sheet columns
sum correctly and the Combined statewide total.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_consolidate_intersection.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import consolidate_intersection_detail as cid
import consolidate_intersection_summary as cis
import outcome
import summary_layout as sl
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


# A trimmed Intersection Detail header (real sheet has 36 cols; the wrapper is
# header-agnostic — it locks whatever the first file has and prepends Route).
_HDR = ["P", "Post Mile", "Location", "Date of Record", "Description"]


def _write_route(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = cid.SHEET_NAME
    ws.append(_HDR)
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


_IS_KEY = {c.slug: c.key for sec in sl.INTERSECTION_SUMMARY_SPEC.sections for c in sec.cats}


def _write_is_route(path, route, total, blocks):
    """A synthetic per-route Intersection Summary sheet (block-count layout)."""
    wb = Workbook()
    ws = wb.active
    ws.title = cis.SHEET_NAME
    ws.append(["TSAR - Intersection Summary"])
    ws.append([f"Route: {route}"])
    ws.append([f"Total Intersections = {total}"])
    ws.append([])
    for hdr, rows in blocks:
        ws.append([hdr])
        ws.append(["NUMBER", "CODE"])
        for cnt, code in rows:
            ws.append([cnt, code])
        ws.append([])
    wb.save(path)
    wb.close()


def _full_blocks(u, d, a, s, total, mastarm_hdr="MAINLINE MASTARM"):
    """A STRUCTURALLY SOUND synthetic route: every non-exempt section sums to
    `total` (the v0.25.0 layout-drift tripwire enforces the partition; only the
    site-under-counted HIGHWAY GROUP is exempt). `mastarm_hdr` lets a fixture
    use the July-2026 MASTERARM spelling (the parse-only alias)."""
    return [
        ("HIGHWAY GROUP", [(u, "U-UNDIVIDED"), (d, "D-DIVIDED")]),
        ("RURAL/URBAN/SUBURBAN", [(total, "R-RURAL -I INSIDE CITY")]),
        ("INTERSECTION TYPE", [(total, "F-FOUR-LEGGED")]),
        ("LIGHTING TYPE", [(total, "N-NO LIGHTING")]),
        ("CONTROL TYPES", [(a, "A-NO CONTROL"), (s, "S-SIGNALIZED")]),
        ("MAINLINE NUM OF LANES", [(total, "2")]),
        (mastarm_hdr, [(total, "Y-YES")]),
        ("MAINLINE LEFT CHANNELIZATION", [(total, "N-NO LEFT TURN CHANNELIZATION")]),
        ("MAINLINE RIGHT CHANNELIZATION", [(total, "N-NO FREE RIGHT TURNS")]),
        ("MAINLINE TRAFFIC FLOW", [(total, "P-2 WAY WITH LEFT TURN")]),
    ]


def test_summary():
    print("Intersection Summary consolidator (block-walk category summer):")
    check("subdir intersection_summary", cis.SUBDIR == "intersection_summary")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_int_sum_"))
    in_dir = tmp / "in"
    in_dir.mkdir()
    # Route 001 uses the July-2026 MASTERARM spelling — the alias must file its
    # counts under the same mastarm categories (all output text stays MASTARM).
    _write_is_route(in_dir / "intersection_summary_route_001.xlsx", "001", 5,
                    _full_blocks(3, 2, 4, 1, 5, mastarm_hdr="MAINLINE MASTERARM"))
    _write_is_route(in_dir / "intersection_summary_route_002.xlsx", "002", 5,
                    _full_blocks(5, 0, 5, 0, 5))
    out = tmp / "out.xlsx"
    res = cis.consolidate(events=Events(), confirm_overwrite=lambda _p: True,
                          input_dir=in_dir, out_path=out)
    check("consolidate ok", res.status == "ok")
    wb = load_workbook(out, read_only=True, data_only=True)
    ws = wb[cis.SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    header = [("" if c is None else str(c)) for c in rows[0]]
    data = rows[1:]
    wb.close()
    check("Route + Total lead the per-route sheet",
          header[0] == "Route" and header[1] == "Total Intersections")
    check("both routes present", {str(r[0]) for r in data} == {"001", "002"})
    u_col = header.index(_IS_KEY["is_highway_group_u"])
    s_col = header.index(_IS_KEY["is_control_types_s"])
    m_col = header.index(_IS_KEY["is_mainline_mastarm_y"])
    u_sum = sum(r[u_col] for r in data)
    s_sum = sum(r[s_col] for r in data)
    m_sum = sum(r[m_col] for r in data)
    check("U-UNDIVIDED column sums across routes (3+5=8)", u_sum == 8)
    check("S-SIGNALIZED column sums across routes (1+0=1)", s_sum == 1)
    check("the MASTERARM-spelled block filed under the MASTARM categories (5+5=10)",
          m_sum == 10)
    cmb = load_workbook(out, data_only=True)[cis.COMBINED_SHEET]
    check("Combined statewide Total = 10", cmb["B2"].value == 10)

    # The v0.25.0 layout-drift tripwire: a route whose (non-exempt) block sums
    # break the route-total partition — here a MISSING mastarm block, the exact
    # shape a silently-renamed header produces — must FAIL that route loudly
    # (named block, PARTIAL completion), never write wrong numbers quietly.
    in2 = tmp / "in2"
    in2.mkdir()
    _write_is_route(in2 / "intersection_summary_route_001.xlsx", "001", 5,
                    _full_blocks(3, 2, 4, 1, 5))
    broken = [b for b in _full_blocks(5, 0, 5, 0, 5)
              if b[0] != "MAINLINE MASTARM"]
    _write_is_route(in2 / "intersection_summary_route_002.xlsx", "002", 5, broken)
    out2 = tmp / "out2.xlsx"
    res2 = cis.consolidate(events=Events(), confirm_overwrite=lambda _p: True,
                           input_dir=in2, out_path=out2)
    check("layout drift -> the bad route FAILS and the result is PARTIAL",
          res2.status == "ok" and res2.completion == outcome.PARTIAL
          and res2.failed_inputs == 1)
    feed = []
    for hdr, rows_ in _full_blocks(3, 2, 4, 1, 5):
        feed.append((None, hdr))
        for cnt, code in rows_:
            feed.append((cnt, code))
    good = sl.counts_from_rows(sl.INTERSECTION_SUMMARY_SPEC, feed)
    check("a sound partition passes the tripwire", cis._layout_drift(good, 5) is None)
    broken_counts = {k: (0 if "mastarm" in k else v) for k, v in good.items()}
    check("the drift names the offending block",
          "MAINLINE MASTARM" in (cis._layout_drift(broken_counts, 5) or ""))


def main():
    print("config (no drift):")
    check("subdir intersection_detail", cid.SUBDIR == "intersection_detail")
    check("sheet name 'Intersection Detail'", cid.SHEET_NAME == "Intersection Detail")
    check("output filename", cid.FILENAME == "tsar_intersection_detail_consolidated.xlsx")

    print("end-to-end 2-route consolidation:")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_int_det_"))
    in_dir = tmp / "in"
    in_dir.mkdir()
    _write_route(in_dir / "intersection_detail_route_001.xlsx",
                 [["R", "0.204", "12 ORA 001", "21-12-31", "JCT 5"],
                  ["", "4.901", "12 ORA 001", "21-12-31", "JCT FOO"]])
    _write_route(in_dir / "intersection_detail_route_002.xlsx",
                 [["R", "1.000", "12 ORA 002", "21-12-31", "BAR ST"]])
    out = tmp / "out.xlsx"
    res = cid.consolidate(events=Events(), confirm_overwrite=lambda _p: True,
                          input_dir=in_dir, out_path=out)
    check("consolidate ok", res.status == "ok")

    wb = load_workbook(out, read_only=True, data_only=True)
    ws = wb[cid.SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [("" if c is None else str(c)) for c in rows[0]]
    data = rows[1:]
    check("leading Route column added", header[0] == "Route")
    check("original header preserved after Route", header[1:6] == _HDR)
    check("all 3 data rows combined", len(data) == 3)
    routes = {str(r[0]) for r in data}
    check("both routes present (001, 002)", routes == {"001", "002"})

    test_summary()

    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL CONSOLIDATE-INTERSECTION CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
