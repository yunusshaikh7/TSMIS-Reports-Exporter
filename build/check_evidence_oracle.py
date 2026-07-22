"""The published comparison, verified against the INDEPENDENT oracle.

CMP-AUD-208's acceptance bar is not "the decoder reads what the engine wrote" —
that only proves the product agrees with itself. It is: *an independent reader
must recompute that cell from immutable sources* and the published cell must
agree. So this check runs both sides over ONE set of raw records:

  raw records ──> phase3_independent_oracle (stdlib only, no product code)
       │                                              │
       └────────> compare_core.run_compare ──> published workbook ──> decoder

and requires them to agree cell for cell, column for column, class for class.
The oracle knows nothing about state masks, workbooks, or openpyxl; the product
knows nothing about the oracle. Anywhere they disagree, the published evidence
is not verified — and the last section proves that disagreement is actually
detected, by corrupting a published mask and requiring the comparison to fail.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_evidence_oracle.py
"""
import os
import shutil
import sys
import tempfile
from collections import Counter
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import phase3_independent_oracle as oracle
import published_comparison as pc
import visual_evidence as ve
from compare_core import CompareSchema, published_key_text, run_compare
from openpyxl import load_workbook

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


# --------------------------------------------------------------------------- #
# ONE corpus, two independent readings
# --------------------------------------------------------------------------- #
COLUMNS = ["PM", "County", "Note", "FT", "Description"]
CONTEXT = ("Note",)
PRODUCT_SCHEMA = CompareSchema(report_name="Oracle", header=COLUMNS,
                               key_field=0, context_fields=CONTEXT)
ORACLE_SCHEMA = oracle.OracleSchema(
    key_rules=(oracle.ValueRule(name="Route"), oracle.ValueRule(name="PM")),
    field_rules=tuple(
        oracle.FieldRule(name=name, asserting=name not in CONTEXT)
        for name in COLUMNS if name != "PM"))
FIELDS = [name for name in COLUMNS if name != "PM"]


class Adapter:
    """The extraction seam: raw record -> OracleRow. No product code."""

    def adapt(self, raw_record, *, side, source_index, schema):
        del side
        return oracle.OracleRow(
            source_index=source_index,
            key=(raw_record["Route"], raw_record["PM"]),
            values=tuple(raw_record[rule.name] for rule in schema.field_rules),
            source_ref=f"row {source_index}")


def raw(route, pm, county="ALA", note="n", ft="1", desc="D"):
    return {"Route": route, "PM": pm, "County": county, "Note": note,
            "FT": ft, "Description": desc}


# Every class the ledger must account for: identical, a counted difference, a
# context-only difference, both one-sided shapes, an unequal-multiplicity
# duplicate group, and a duplicate group whose differences exist only inside it.
RAW_A = [
    raw("001", "1.000"),
    raw("001", "2.000", desc="ALPHA"),
    raw("001", "2.500", ft="7"),
    raw("001", "3.000", note="left"),
    raw("001", "4.000"),
    raw("002", "9.000", desc="DUP-A1"),
    raw("002", "9.000", desc="DUP-A2"),
    raw("003", "5.000", desc="TRIPLE-A1"),
    raw("003", "5.000", desc="TRIPLE-A2"),
    raw("003", "5.000", desc="TRIPLE-A3"),
]
RAW_B = [
    raw("001", "1.000"),
    raw("001", "2.000", desc="BETA"),
    raw("001", "2.500", ft="9"),
    raw("001", "3.000", note="right"),
    raw("001", "5.000"),
    raw("002", "9.000", desc="DUP-B1"),
    raw("002", "9.000", desc="DUP-B2"),
    raw("003", "5.000", desc="TRIPLE-A1"),
    raw("003", "5.000", desc="TRIPLE-B2"),
]


def product_rows(records):
    return [[r["Route"]] + [r[name] for name in COLUMNS] for r in records]


tmp = Path(tempfile.mkdtemp(prefix="check_ev_oracle_"))
values = tmp / "cmp.xlsx"
result = run_compare(PRODUCT_SCHEMA, product_rows(RAW_A), product_rows(RAW_B),
                     True, values, mode="values",
                     confirm_overwrite=lambda _p: True)
outcome = oracle.compare_raw_records(ORACLE_SCHEMA, RAW_A, RAW_B, Adapter())
published = pc.read(values)
ledger = published.ledger()

print("both readings completed")
check("the product published a comparison", result.status == "ok")
check("the oracle produced a complete outcome",
      outcome.completion == "complete" and outcome.pairing_quality == "exact")
check("the oracle sees differences (a vacuous corpus proves nothing)",
      outcome.counts.differing_cells > 0)

# --------------------------------------------------------------------------- #
print("aggregate agreement")
check("counted difference cells agree",
      ledger.difference_cells == outcome.counts.differing_cells)
check("per-column difference counts agree",
      {name: ledger.differences(name) for name in FIELDS
       if ledger.differences(name)}
      == dict(outcome.counts.per_field_counts))
check("matched-row counts agree",
      ledger.matched_rows == outcome.counts.paired_rows)
check("one-sided row counts agree per side",
      dict(ledger.one_sided_by_status)
      == {"TSMIS only": outcome.counts.side_a_only_rows,
          "TSN only": outcome.counts.side_b_only_rows})
check("context cells agree", ledger.context_cells == outcome.counts.context_cells)
check("asserted cells agree",
      ledger.equal_cells + ledger.difference_cells
      == outcome.counts.asserted_cells)
check("the row universe agrees",
      ledger.data_rows == (outcome.counts.paired_rows
                           + outcome.counts.side_a_only_rows
                           + outcome.counts.side_b_only_rows))

# --------------------------------------------------------------------------- #
print("cell-for-cell agreement")
# Address each oracle row result by the identity the product published for its
# side-A source row; the oracle carries the stable source indices.
product_a = product_rows(RAW_A)
oracle_by_address = {}
for row_result in outcome.row_results:
    source = product_a[row_result.source_index_a]
    key = published_key_text(PRODUCT_SCHEMA, source)
    address = (source[0], key)
    oracle_by_address.setdefault(address, []).append(row_result)

matched = mismatched = unaddressed = 0
for address, results in oracle_by_address.items():
    rows = [r for r in published.rows
            if (r.route, r.key) == address and r.matched]
    if len(rows) != len(results):
        unaddressed += 1
        continue
    # Within a duplicate group both sides publish the same multiset of cell
    # states; compare as multisets so occurrence numbering is not asserted here
    # (the pairing objective owns WHICH occurrence pairs with which).
    published_masks = Counter(r.mask for r in rows)
    oracle_masks = Counter()
    for row_result in results:
        oracle_masks["".join(
            "D" if cell.counts_as_difference
            else "N" if not cell.asserting else "E"
            for cell in row_result.cells)] += 1
    if published_masks == oracle_masks:
        matched += len(rows)
    else:
        mismatched += len(rows)
check("every matched row's published state mask equals the oracle's",
      mismatched == 0 and unaddressed == 0 and matched == ledger.matched_rows)
check("one-sided rows publish an all-U mask on both sides",
      all(set(r.mask) == {"U"} for r in published.rows if not r.matched))

# --------------------------------------------------------------------------- #
print("the evidence spine agrees with the oracle")
# A candidate is renderable only for a key with exactly one row on each side.
solo_addresses = {
    address for address, results in oracle_by_address.items()
    if len(results) == 1
    and len([r for r in published.rows if (r.route, r.key) == address]) == 1}
oracle_renderable = Counter()
for address in solo_addresses:
    for cell, rule in zip(oracle_by_address[address][0].cells,
                          ORACLE_SCHEMA.field_rules):
        if cell.counts_as_difference:
            oracle_renderable[rule.name] += 1
check("the ledger's unique-row differences equal the oracle's renderable set",
      {name: ledger.solo_differences(name) for name in FIELDS
       if ledger.solo_differences(name)} == dict(oracle_renderable))
check("the ledger's duplicate-group differences are the oracle's remainder",
      {name: ledger.duplicate_differences(name) for name in FIELDS
       if ledger.duplicate_differences(name)}
      == {name: outcome.counts.per_field_counts[name]
          - oracle_renderable.get(name, 0)
          for name in outcome.counts.per_field_counts
          if outcome.counts.per_field_counts[name]
          - oracle_renderable.get(name, 0)})

candidates = {}
for address in sorted(solo_addresses):
    row = [r for r in published.rows if (r.route, r.key) == address][0]
    for cell, rule in zip(oracle_by_address[address][0].cells,
                          ORACLE_SCHEMA.field_rules):
        if not cell.counts_as_difference:
            continue
        position = published.position_of(rule.name)
        candidates.setdefault(rule.name, []).append(dict(
            route=address[0], key=address[1], field=rule.name,
            va=cell.normalized_a.text, vb=cell.normalized_b.text,
            dist="", cnty="", pub_key=address[1],
            display=row.value(position)))
kept, rejected = ve._reconcile(published, candidates)
check("every oracle-derived renderable cell reconciles to its published cell",
      not rejected
      and {name: len(items) for name, items in kept.items()}
      == dict(oracle_renderable))
check("each reconciled item names its published row and state",
      all(item["published_state"] == "D" and item["published_row"] > 1
          for items in kept.values() for item in items))

# --------------------------------------------------------------------------- #
print("the agreement test has teeth")
tampered = tmp / "tampered.xlsx"
shutil.copy2(values, tampered)
wb = load_workbook(tampered)
ws = wb["Comparison"]
state_col = [c.column for c in ws[1]
             if isinstance(c.value, str) and "_STATE_V" in c.value][0]
diffs_col = [c.column for c in ws[1] if c.value == "Diffs"][0]
flipped = None
for r in range(2, ws.max_row + 1):
    mask = ws.cell(row=r, column=state_col).value
    if isinstance(mask, str) and "D" in mask:
        flipped = (r, mask)
        ws.cell(row=r, column=state_col, value=mask.replace("D", "E", 1))
        ws.cell(row=r, column=diffs_col,
                value=int(ws.cell(row=r, column=diffs_col).value) - 1)
        break
wb.save(tampered)
check("a difference cell was available to tamper with", flipped is not None)
tampered_ledger = pc.read(tampered).ledger()
check("a downgraded published cell no longer agrees with the oracle",
      tampered_ledger.difference_cells != outcome.counts.differing_cells)
check("...and the tampered ledger's digest differs from the honest one",
      tampered_ledger.digest() != ledger.digest())

# The oracle must also be sensitive in the other direction: change the SOURCE
# and the two readings must move together, never one alone.
moved_a = [dict(record) for record in RAW_A]
moved_a[0]["Description"] = "CHANGED"
moved_values = tmp / "moved.xlsx"
run_compare(PRODUCT_SCHEMA, product_rows(moved_a), product_rows(RAW_B), True,
            moved_values, mode="values", confirm_overwrite=lambda _p: True)
moved_outcome = oracle.compare_raw_records(ORACLE_SCHEMA, moved_a, RAW_B,
                                           Adapter())
moved_ledger = pc.read(moved_values).ledger()
check("a source edit moves the published count and the oracle count together",
      moved_ledger.difference_cells == moved_outcome.counts.differing_cells
      == outcome.counts.differing_cells + 1)

shutil.rmtree(tmp, ignore_errors=True)

print()
if _fail:
    print(f"FAILED {len(_fail)} check(s):")
    for name in _fail:
        print(f"  - {name}")
    sys.exit(1)
print("evidence oracle: all checks passed")
