"""Golden check for the Compare-tab "vs Baseline" matrix: the baseline_matrix
engine (baseline identity parsing, available days + picker options, snapshot
cell states incl. the baseline's own column + fingerprint staleness, the scoped
rebuild list, build_baseline_cell guard paths AND one REAL end-to-end build via
compare_env — both baseline kinds, with the explicit side labels) plus the
gui_api bridge (source/baseline/day/report validation, enqueue onto the SHARED
matrix queue with which="baseline", open guards). The real build runs on tiny
synthetic Ramp Detail fixtures — no browser, no TSN.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_baseline_matrix.py
"""
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import Workbook, load_workbook

import baseline_matrix
import gui_api
import gui_matrix
import paths
import settings

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


class _FakeWorker:
    last = None

    def __init__(self, *args, **kwargs):
        _FakeWorker.last = (args, kwargs)

    def start(self):
        self.started = True


def _touch(p, data=b"PK"):
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(data)


def _rd_route_file(p, desc):
    """A per-route Ramp Detail export with the REAL 11-column layout + a PM-keyed
    row. CMP-AUD-032 pins the cross-env schema, so a short fake header is now
    refused — the fixture must carry the true site header."""
    import compare_ramp_detail_tsn as _rd
    wb = Workbook()
    ws = wb.active
    ws.title = "TSAR - Ramp Detail"
    hdr = list(_rd._TSMIS_HEADER[1:])
    ws.append(hdr)
    row = [""] * len(hdr)
    row[0] = "01-DN-101"                      # Location
    row[hdr.index("PM")] = "1.000"
    row[hdr.index("Description")] = desc
    ws.append(row)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


def _raises(fn):
    try:
        fn()
        return False
    except ValueError:
        return True


def main():
    out = Path(tempfile.mkdtemp(prefix="tsmis_bl_out_"))
    dest = Path(tempfile.mkdtemp(prefix="tsmis_bl_dest_"))
    cfgdir = Path(tempfile.mkdtemp(prefix="tsmis_bl_cfg_"))
    saved = (paths.OUTPUT_ROOT, baseline_matrix.OUTPUT_ROOT,
             gui_matrix.BaselineMatrixCompareWorker, settings.get_batch_dest,
             settings.CONFIG_FILE)
    paths.OUTPUT_ROOT = out
    baseline_matrix.OUTPUT_ROOT = out
    gui_matrix.BaselineMatrixCompareWorker = _FakeWorker
    settings.get_batch_dest = lambda: str(dest)
    settings.CONFIG_FILE = cfgdir / "config.json"
    settings._cache = settings._cache_mtime = None
    try:
        print("baseline identity — parse, token, dir, labels:")
        check("store id parses", baseline_matrix.parse_baseline("store") == ("store", None))
        check("day id parses",
              baseline_matrix.parse_baseline("day:2026-06-11") == ("day", "2026-06-11"))
        check("unset/garbage/traversal ids rejected",
              baseline_matrix.parse_baseline(None) is None
              and baseline_matrix.parse_baseline("") is None
              and baseline_matrix.parse_baseline("day:..\\evil") is None
              and baseline_matrix.parse_baseline("day:2026-06-11/../x") is None)
        check("tokens are filename-safe",
              baseline_matrix.baseline_token("store") == "store"
              and baseline_matrix.baseline_token("day:2026-06-11") == "2026-06-11")
        check("baseline dirs resolve (day -> run folder, store -> dest/<source>)",
              baseline_matrix.baseline_dir("ssor-prod", "day:2026-06-11", str(dest))
              == out / "2026-06-11 ssor-prod"
              and baseline_matrix.baseline_dir("ssor-prod", "store", str(dest))
              == dest / "ssor-prod")
        check("labels distinct per kind",
              baseline_matrix.baseline_label("ssor-prod", "store") == "SSOR-PROD (store)"
              and baseline_matrix.baseline_label("ssor-prod", "day:2026-06-11")
              == "SSOR-PROD 2026-06-11")

        # Plant two ssor-prod days (Ramp Detail real fixtures; Highway Log stub)
        # + a store copy for the same source.
        _rd_route_file(out / "2026-06-18 ssor-prod" / "ramp_detail"
                       / "tsar_ramp_detail_route_101.xlsx", "RAMP A NEW")
        _rd_route_file(out / "2026-06-11 ssor-prod" / "ramp_detail"
                       / "tsar_ramp_detail_route_101.xlsx", "RAMP A OLD")
        _rd_route_file(dest / "ssor-prod" / "ramp_detail"
                       / "tsar_ramp_detail_route_101.xlsx", "RAMP A NEW")
        _touch(out / "2026-06-18 ssor-prod" / "highway_log" / "r1.xlsx")

        print("available days + baseline options:")
        check("days scoped to the source, newest first",
              baseline_matrix.available_days("ssor-prod") == ["2026-06-18", "2026-06-11"]
              and baseline_matrix.available_days("ars-prod") == [])
        opts = baseline_matrix.baseline_options("ssor-prod", str(dest))
        check("options = store + both days, with per-option report coverage",
              [o["id"] for o in opts] == ["store", "day:2026-06-18", "day:2026-06-11"]
              and opts[0]["present"] == 1 and opts[1]["present"] == 2
              and opts[2]["present"] == 1 and opts[0]["total"] == 13)

        print("snapshot — cell states:")
        days = ["2026-06-18", "2026-06-11"]
        snap0 = baseline_matrix.baseline_matrix_snapshot(
            "ssor-prod", days, None, dest=str(dest))
        check("13 report rows, all supported",
              len(snap0["all_rows"]) == 13
              and all(r["supported"] for r in snap0["all_rows"]))
        check("no baseline picked -> baseline side missing on every cell",
              snap0["baseline"]["id"] is None
              and snap0["cells"]["ramp_detail"]["2026-06-18"]["cmp"]["missing_side"]
              == "baseline")
        snap = baseline_matrix.baseline_matrix_snapshot(
            "ssor-prod", days, "day:2026-06-11", dest=str(dest))
        check("baseline meta resolved (kind/date/label/per-row presence)",
              snap["baseline"]["kind"] == "day" and snap["baseline"]["date"] == "2026-06-11"
              and snap["baseline"]["present"]["ramp_detail"]["present"] is True
              and snap["baseline"]["present"]["highway_log"]["present"] is False)
        cell = snap["cells"]["ramp_detail"]["2026-06-18"]
        check("comparable RD cell: both sides present, not built",
              cell["cmp"]["missing_side"] is None and not cell["cmp"]["built"])
        check("the baseline's own column renders is_baseline",
              snap["cells"]["ramp_detail"]["2026-06-11"]["cmp"].get("is_baseline") is True)
        check("HL cell: export present but baseline copy missing",
              snap["cells"]["highway_log"]["2026-06-18"]["cmp"]["missing_side"]
              == "baseline")

        print("scoped rebuild list:")
        todo = baseline_matrix.cells_to_rebuild(snap, scope="all")
        check("all-scope = the one comparable RD cell (baseline column + "
              "missing sides skipped)",
              todo == [("2026-06-18", "ramp_detail")])
        check("row/date filters scope",
              baseline_matrix.cells_to_rebuild(snap, "all", row="highway_log") == []
              and baseline_matrix.cells_to_rebuild(snap, "all", date="2026-06-11") == [])

        print("build_baseline_cell — guards + one REAL build per baseline kind:")
        check("unknown row raises",
              _raises(lambda: baseline_matrix.build_baseline_cell(
                  "ssor-prod", "2026-06-18", "nope", "day:2026-06-11", str(dest), None)))
        check("bad date raises",
              _raises(lambda: baseline_matrix.build_baseline_cell(
                  "ssor-prod", "..", "ramp_detail", "day:2026-06-11", str(dest), None)))
        check("unset baseline raises",
              _raises(lambda: baseline_matrix.build_baseline_cell(
                  "ssor-prod", "2026-06-18", "ramp_detail", "", str(dest), None)))
        check("the baseline's own day raises",
              _raises(lambda: baseline_matrix.build_baseline_cell(
                  "ssor-prod", "2026-06-11", "ramp_detail", "day:2026-06-11",
                  str(dest), None)))
        res = baseline_matrix.build_baseline_cell(
            "ssor-prod", "2026-06-18", "ramp_detail", "day:2026-06-11", str(dest), None)
        outwb = baseline_matrix.out_path("2026-06-18", "ssor-prod", "ramp_detail",
                                         "day:2026-06-11")
        check("day-baseline build ok, workbook written under baseline-by-day",
              res.status == "ok" and outwb.is_file()
              and outwb.parent == baseline_matrix.byday_root() / "2026-06-18 ssor-prod")
        wb = load_workbook(outwb, read_only=True)
        names = set(wb.sheetnames)
        wb.close()
        check("side labels carry the day + baseline identities",
              "Only in SSOR-PROD 2026-06-18" in names
              and "Only in SSOR-PROD 2026-06-11" in names)
        rec = baseline_matrix.load_results().get(
            "2026-06-18 ssor-prod|ramp_detail|day:2026-06-11")
        check("counts cached under the baseline-qualified key (1 diff cell)",
              rec is not None and rec["diff_cells"] == 1 and rec["one_sided"] == 0)
        res2 = baseline_matrix.build_baseline_cell(
            "ssor-prod", "2026-06-18", "ramp_detail", "store", str(dest), None)
        outwb2 = baseline_matrix.out_path("2026-06-18", "ssor-prod", "ramp_detail",
                                          "store")
        wb = load_workbook(outwb2, read_only=True)
        names2 = set(wb.sheetnames)
        wb.close()
        check("store-baseline build ok, its own artifact, '(store)' label",
              res2.status == "ok" and outwb2.is_file() and outwb2 != outwb
              and "Only in SSOR-PROD (store)" in names2)
        rec2 = baseline_matrix.load_results().get(
            "2026-06-18 ssor-prod|ramp_detail|store")
        check("store build identical (same-content sides -> 0 diffs)",
              rec2 is not None and rec2["diff_cells"] == 0)

        # v0.38.3: evidence images on the vs-Baseline matrix. Both sides of a
        # cell here are the SAME report's per-route prints from two folders, so
        # the request rides visual_evidence's env flavor unchanged. These pin the
        # WIRING (the render itself is the env lane's own, already locked in
        # check_visual_evidence); the real print render is exercised end to end
        # against the corpus, not from synthetic fixtures.
        print("evidence wiring (v0.38.3) — the env lane, day-picked:")
        import visual_evidence as _ve
        import matrix as _mx
        seen = []
        _real_run_env_evidence = _mx._run_env_evidence
        _mx._run_env_evidence = lambda *a: seen.append(a)
        try:
            baseline_matrix.build_baseline_cell(
                "ssor-prod", "2026-06-18", "ramp_detail", "day:2026-06-11",
                str(dest), None,
                evidence={"enabled": True, "examples": 3, "layout": "pair"})
            check("a successful build forwards the evidence request with the "
                  "cell's row key and its OWN committed workbook",
                  len(seen) == 1 and seen[0][0] == "ramp_detail"
                  and Path(seen[0][1]) == outwb
                  and seen[0][2] == {"enabled": True, "examples": 3,
                                     "layout": "pair"})
            seen.clear()
            baseline_matrix.build_baseline_cell(
                "ssor-prod", "2026-06-18", "ramp_detail", "day:2026-06-11",
                str(dest), None)
            check("no evidence request -> the decoration is still called and "
                  "no-ops on None (one path, not two)",
                  len(seen) == 1 and seen[0][2] is None)
        finally:
            _mx._run_env_evidence = _real_run_env_evidence
        # With the REAL decoration, an EXCEL row must publish nothing: evidence
        # is print crops, and an Excel export has no print to crop.
        baseline_matrix.build_baseline_cell(
            "ssor-prod", "2026-06-18", "ramp_detail", "day:2026-06-11",
            str(dest), None, evidence={"enabled": True, "examples": 2})
        check("an EXCEL row generates NOTHING even with the toggle on "
              "(env_capable refuses it; no evidence sibling appears)",
              not _ve.env_capable("ramp_detail")
              and not _ve.sibling_paths(outwb)[0].exists())
        check("every `_pdf` row IS capable on this lane — Highway Detail "
              "included (the row the owner asked for)",
              all(_ve.env_capable(rk) for rk in
                  ("highway_detail_pdf", "highway_log_pdf", "highway_sequence_pdf",
                   "intersection_detail_pdf", "ramp_detail_pdf")))
        check("run_baseline_evidence_only refuses an Excel row, an unbuilt "
              "`_pdf` cell, and the baseline's own day",
              _raises(lambda: baseline_matrix.run_baseline_evidence_only(
                  "ssor-prod", "2026-06-18", "ramp_detail", "day:2026-06-11",
                  str(dest), None))
              and _raises(lambda: baseline_matrix.run_baseline_evidence_only(
                  "ssor-prod", "2026-06-18", "ramp_detail_pdf", "day:2026-06-11",
                  str(dest), None))
              and _raises(lambda: baseline_matrix.run_baseline_evidence_only(
                  "ssor-prod", "2026-06-11", "highway_detail_pdf",
                  "day:2026-06-11", str(dest), None)))

        print("freshness — mtime + input identity (both folders):")
        snap2 = baseline_matrix.baseline_matrix_snapshot(
            "ssor-prod", days, "day:2026-06-11", dest=str(dest))
        c2 = snap2["cells"]["ramp_detail"]["2026-06-18"]["cmp"]
        check("built cell reads fresh with cached counts",
              c2["built"] and not c2["stale"] and c2["diff_cells"] == 1)
        # A route DELETED from the BASELINE side is invisible to newest-mtime —
        # the recorded two-folder fingerprint must flag the cell stale.
        extra = (out / "2026-06-11 ssor-prod" / "ramp_detail"
                 / "tsar_ramp_detail_route_202.xlsx")
        _rd_route_file(extra, "SECOND ROUTE")
        baseline_matrix.build_baseline_cell(
            "ssor-prod", "2026-06-18", "ramp_detail", "day:2026-06-11", str(dest), None)
        extra.unlink()
        snap3 = baseline_matrix.baseline_matrix_snapshot(
            "ssor-prod", days, "day:2026-06-11", dest=str(dest))
        check("a route deleted on the baseline side reads the cell STALE "
              "(fingerprint, not mtime)",
              snap3["cells"]["ramp_detail"]["2026-06-18"]["cmp"]["stale"] is True)

        print("gui_api bridge — validation + enqueue (shared queue):")
        a = gui_api.GuiApi()
        info = a.baseline_matrix_info()
        check("baseline_matrix_info carries available_days + baseline_options",
              info["available_days"] == ["2026-06-18", "2026-06-11"]
              and [o["id"] for o in info["baseline_options"]][0] == "store")
        check("unknown source rejected",
              bool(a.set_baseline_matrix_source("zz-zz").get("error")))
        check("valid source set", a.set_baseline_matrix_source("ssor-prod").get("ok"))
        check("an export-less baseline id rejected",
              bool(a.set_baseline_matrix_baseline("day:2099-01-01").get("error")))
        check("valid baseline persisted",
              a.set_baseline_matrix_baseline("day:2026-06-11").get("ok")
              and settings.get_baseline_matrix_baseline() == "day:2026-06-11")
        check("clearing the baseline ok",
              a.set_baseline_matrix_baseline("").get("ok")
              and settings.get_baseline_matrix_baseline() == "")
        a.set_baseline_matrix_baseline("day:2026-06-11")
        check("add a day with no export rejected",
              bool(a.add_baseline_matrix_day("2099-01-01").get("error")))
        check("add a real day", a.add_baseline_matrix_day("2026-06-18").get("ok")
              and "2026-06-18" in settings.get_baseline_matrix_days())
        a.add_baseline_matrix_day("2026-06-11")
        # CMP-AUD-095: switching source reconciles BOTH the retained day columns and
        # the baseline id — ars-prod has no exports here, so both must clear rather
        # than aim a build at folders that don't exist for the new source.
        a.set_baseline_matrix_source("ars-prod")
        check("source switch clears source-scoped day columns",
              settings.get_baseline_matrix_days() == [])
        check("source switch clears a baseline invalid for the new source",
              settings.get_baseline_matrix_baseline() == "")
        a.set_baseline_matrix_source("ssor-prod")        # restore for downstream
        settings.set_baseline_matrix_days(["2026-06-18", "2026-06-11"])
        settings.set_baseline_matrix_baseline("day:2026-06-11")
        check("hide unknown report rejected",
              bool(a.set_baseline_matrix_report("nope", False).get("error")))
        check("build on the baseline's own day rejected",
              bool(a.build_baseline_matrix_cell("ramp_detail", "2026-06-11").get("error")))
        bc = a.build_baseline_matrix_cell("ramp_detail", "2026-06-18")
        check("build a comparable cell -> launched as a 'matrix' task via the "
              "baseline worker",
              bc.get("ok") is True and a._task == "matrix"
              and isinstance(_FakeWorker.last, tuple)
              and _FakeWorker.last[0][2] == "day:2026-06-11")
        q = a.build_baseline_matrix_cell("ramp_detail", "2026-06-18")
        check("second action queues behind it (shared queue)",
              q.get("ok") is True and len(a._state_snapshot()["matrix_queue"]) == 1)
        a._end_task()
        a._end_task()
        check("queue drained -> idle", a._task is None)
        settings.set_baseline_matrix_baseline("")
        check("rebuild without a baseline errors",
              bool(a.rebuild_baseline_matrix("all").get("error")))
        settings.set_baseline_matrix_baseline("day:2026-06-11")
        rb = a.rebuild_baseline_matrix("all")
        check("rebuild-all launches over the comparable cells",
              rb.get("ok") and a._task == "matrix")
        a._end_task()
        # CMP-AUD-096: a supplied-but-invalid row/date is REJECTED, not widened to all.
        check("rebuild_baseline_matrix rejects a supplied-but-invalid row (CMP-AUD-096)",
              bool(a.rebuild_baseline_matrix("all", row="nope").get("error")))
        check("rebuild_baseline_matrix rejects an impossible date (CMP-AUD-096)",
              bool(a.rebuild_baseline_matrix("all", date="2026-99-99").get("error")))
        check("rebuild_baseline_matrix stays idle after a rejected scope", a._task is None)
        check("state carries the baseline formulas toggle",
              a._state_snapshot().get("baseline_matrix_formulas") is False
              and a.set_baseline_matrix_formulas(True).get("on") is True
              and a._state_snapshot().get("baseline_matrix_formulas") is True)

        # v0.38.3: the per-cell camera. The job must reach the BASELINE
        # resolver — dispatching it to the Everything matrix's env resolver
        # would illustrate a different comparison entirely.
        check("evidence camera refuses an Excel row",
              bool(a.baseline_matrix_evidence_cell("ramp_detail",
                                                   "2026-06-18").get("error")))
        check("evidence camera refuses the baseline's own day",
              bool(a.baseline_matrix_evidence_cell("highway_detail_pdf",
                                                   "2026-06-11").get("error")))
        _saved_ev_worker = gui_matrix.MatrixEvidenceWorker
        gui_matrix.MatrixEvidenceWorker = _FakeWorker
        _FakeWorker.last = None
        try:
            ec = a.baseline_matrix_evidence_cell("highway_detail_pdf", "2026-06-18")
            job = a._state_snapshot().get("matrix_current") or {}
            check("Highway Detail (PDF) camera enqueues an evidence job on the "
                  "SHARED queue tagged which='baseline'",
                  ec.get("ok") is True and job.get("kind") == "evidence"
                  and job.get("which") == "baseline")
            reached = {}
            _saved_only = baseline_matrix.run_baseline_evidence_only
            baseline_matrix.run_baseline_evidence_only = (
                lambda *args, **kw: reached.update(args=args, kw=kw))
            try:
                _FakeWorker.last[0][0](None)      # the dispatched run_fn
            finally:
                baseline_matrix.run_baseline_evidence_only = _saved_only
            check("...and its run_fn targets run_baseline_evidence_only with "
                  "the cell's own source/day/row/baseline (not the env lane)",
                  reached.get("args", ())[:4]
                  == ("ssor-prod", "2026-06-18", "highway_detail_pdf",
                      "day:2026-06-11"))
        finally:
            gui_matrix.MatrixEvidenceWorker = _saved_ev_worker
            a._end_task()

        print("gui_api bridge — open guards:")
        opened = []
        a._open_file = lambda p: opened.append(Path(p))
        a._open_folder = lambda p: opened.append(Path(p))
        check("open a never-built cell errors",
              bool(a.open_baseline_cell_comparison("highway_log", "2026-06-18")
                   .get("error")))
        check("open succeeds for the built RD cell",
              a.open_baseline_cell_comparison("ramp_detail", "2026-06-18")
              .get("ok") is True and opened[-1] == baseline_matrix.out_path(
                  "2026-06-18", "ssor-prod", "ramp_detail", "day:2026-06-11"))
        check("open vs-baseline folder ok",
              a.open_baseline_comparisons_folder().get("ok") is True
              and opened[-1] == baseline_matrix.byday_root())
    finally:
        (paths.OUTPUT_ROOT, baseline_matrix.OUTPUT_ROOT,
         gui_matrix.BaselineMatrixCompareWorker, settings.get_batch_dest,
         settings.CONFIG_FILE) = saved
        settings._cache = settings._cache_mtime = None
        for d in (out, dest, cfgdir):
            shutil.rmtree(d, ignore_errors=True)

    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL BASELINE-MATRIX CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
