"""Regression lock for compare_core's composite-style cache.

`_styled_cell` registers each distinct (font, fill, alignment, border) set once
and copies the resulting StyleArray onto every later cell that wants it, instead
of re-walking openpyxl's style descriptors 1.34 million times per statewide
comparison.  That is only allowed to be a SPEED change: this check A/Bs it
against the historical assignment sequence and requires every stable OOXML
package member to be byte-exact.  docProps/core.xml is excluded because its
created/modified timestamps are intentionally run-specific.

It also holds the three properties that make the cache safe:
  * cells own their StyleArray, so a later `number_format` / fill mutation
    cannot reach the prototype or an already-appended sibling;
  * the cache is keyed by component IDENTITY and retains its components, so an
    equal-but-distinct object gets its own entry rather than a recycled id;
  * style is bound BEFORE the value, because openpyxl sets `number_format`
    while binding a date and a later style copy would erase it.
"""
from __future__ import annotations

import sys
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import Cell, WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT / "build"), str(ROOT)]

import compare_core as core  # noqa: E402
from benchmark_vs_tsn_speed import _package_manifest  # noqa: E402


_failures = []


def check(name, condition, detail=""):
    print(f"  [{'OK ' if condition else 'FAIL'}] {name}")
    if not condition:
        _failures.append(name)
        if detail:
            print(f"       {detail}")


def _historical_styled_cell(ws, value, *, font, fill=None, align=None, border=None,
                            guard=False, exact_source_numeric=False):
    """The pre-cache construction sequence, kept here as the A/B control."""
    c = WriteOnlyCell(ws, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    if guard:
        core.set_safe_literal_cell(
            c, value, exact_source_numeric=exact_source_numeric)
    return c


def _ditto_resolver(_rows, _has_route):
    # Exercises the post-construction fill mutation that requires each cell to
    # own its StyleArray instead of aliasing the cached prototype.
    return {0: {2: "paired value"}}


SCHEMA = core.CompareSchema(
    report_name="Style Cache Lock",
    header=["Loc", "Text", "Ditto", "Med Wid", "Context"],
    side_a="TSMIS",
    side_b="TSN",
    id_noun="row",
    id_noun_plural="rows",
    medwid_fields=("Med Wid",),
    context_fields=("Context",),
    ditto_nonasserting=True,
    ditto_resolver=_ditto_resolver,
)

LEFT = [
    ["A", "=1+1", "++", "8V", "left context"],
    ["B", "@literal", "same", "02", ""],
    ["ONLY-A", "#N/A", "+", "0", "context"],
]
RIGHT = [
    ["A", "=1+1", "++", "08V", "right context"],
    ["B", "different", "same", "2", ""],
    ["ONLY-B", "-literal", "+", "0", "context"],
]


def _binding_probe():
    """How many times a guarded literal is bound, and its final data type."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Probe")
    calls = []
    original = Cell._bind_value

    def counted(cell, value):
        calls.append(value)
        return original(cell, value)

    Cell._bind_value = counted
    try:
        cell = core._styled_cell(
            ws, "=1+1", font=Font(name="Arial", size=10), guard=True)
    finally:
        Cell._bind_value = original
        wb.close()
    return calls, cell.data_type


def _hot_loop_contract():
    print("style cache identity + per-cell ownership:")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Identity")
    font = Font(name="Arial", size=10)
    same_value_distinct_font = Font(name="Arial", size=10)
    first = core._styled_cell(ws, "first", font=font)
    second = core._styled_cell(ws, "second", font=font)
    cache = wb._tsmis_style_cache
    first_key = next(iter(cache))
    first_entry = cache[first_key]
    check("cache keys are component identities, and components are retained",
          len(cache) == 1 and all(isinstance(item, int) for item in first_key)
          and first_entry[0] is font and first._style == second._style)
    third = core._styled_cell(ws, "third", font=same_value_distinct_font)
    check("equal-but-distinct components get their own entry and converge",
          font == same_value_distinct_font and len(cache) == 2
          and first._style == third._style)

    # The mutation the Highway Log ditto tint and the Summary grid perform.
    fill = PatternFill("solid", start_color="FFFF00")
    a = core._styled_cell(ws, "a", font=font, fill=fill)
    b = core._styled_cell(ws, "b", font=font, fill=fill)
    a.number_format = "0.000"
    check("a later number_format reaches that cell only",
          a.number_format == "0.000" and b.number_format == "General"
          and core._styled_cell(ws, "c", font=font,
                                fill=fill).number_format == "General")
    wb.close()

    calls, data_type = _binding_probe()
    check("a guarded literal is bound exactly once and stays text",
          calls == ["=1+1"] and data_type == "s", f"binds={calls!r}")

    # openpyxl sets number_format while binding a date: style must go on first.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Dates")
    dated = core._styled_cell(ws, datetime(2026, 8, 20, 9, 30),
                              font=font, align=Alignment(horizontal="center"))
    control = _historical_styled_cell(ws, datetime(2026, 8, 20, 9, 30),
                                      font=font,
                                      align=Alignment(horizontal="center"))
    check("a date keeps its number format (style is bound before the value)",
          dated.number_format == control.number_format != "General",
          f"cached={dated.number_format!r} historical={control.number_format!r}")
    wb.close()


def _one_mode(root, mode):
    control = root / f"{mode}-historical.xlsx"
    cached = root / f"{mode}-cached.xlsx"
    saved = core._styled_cell
    core._styled_cell = _historical_styled_cell
    try:
        control_result = core.run_compare(SCHEMA, LEFT, RIGHT, False, control,
                                          mode=mode)
    finally:
        core._styled_cell = saved
    cached_result = core.run_compare(SCHEMA, LEFT, RIGHT, False, cached, mode=mode)

    check(f"{mode}: both writers build",
          control_result.status == cached_result.status == "ok"
          and control.is_file() and cached.is_file(),
          f"historical={control_result.status!r} cached={cached_result.status!r}")
    if not control.is_file() or not cached.is_file():
        return

    control_typed = control_result.comparison_outcome.to_dict()
    cached_typed = cached_result.comparison_outcome.to_dict()
    check(f"{mode}: typed outcome is exact",
          control_typed == cached_typed,
          f"historical={control_typed!r} cached={cached_typed!r}")

    control_package = _package_manifest(control)
    cached_package = _package_manifest(cached)
    check(f"{mode}: every stable OOXML member is byte-exact",
          control_package == cached_package,
          f"historical={control_package['sha256']} cached={cached_package['sha256']}")
    check(f"{mode}: fixture covers guarded literals and per-cell style mutation",
          control_typed["counts"]["differing_cells"] > 0
          and control_package["member_count"] >= 15)


def _report_view_writers():
    """The two report views build their own bordered cells through the cache."""
    print("report-view writers reach the same cache:")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("RV")
    font = Font(name="Arial", size=9)
    fill = PatternFill("solid", start_color="EEEEEE")
    border = Border(bottom=Side(style="thin"))
    # "=X" is the formula-injection guard; True and the 17-digit float are the
    # two paths that also force number_format="@".
    for value in ("=X", True, 1.2345678901234567e-5, "plain", 42):
        cached = core._styled_cell(ws, value, font=font, fill=fill,
                                   align=Alignment(horizontal="left"),
                                   border=border, guard=True)
        control = _historical_styled_cell(ws, value, font=font, fill=fill,
                                          align=Alignment(horizontal="left"),
                                          border=border, guard=True)
        check(f"bordered guarded report-view cell matches historical: {value!r}",
              cached._style == control._style
              and cached.value == control.value
              and cached.data_type == control.data_type
              and cached.number_format == control.number_format,
              f"cached=({cached.value!r}, {cached.data_type}, "
              f"{cached.number_format!r}) historical=({control.value!r}, "
              f"{control.data_type}, {control.number_format!r})")
    check("the guard fixture really exercises the text-forcing path",
          core._styled_cell(ws, True, font=font, guard=True).number_format == "@")
    wb.close()


def main():
    print("compare_core composite-style cache:")
    _hot_loop_contract()
    _report_view_writers()
    print("byte-for-byte A/B against the historical assignment sequence:")
    with tempfile.TemporaryDirectory(prefix="tsmis_style_cache_") as tmp:
        root = Path(tmp)
        _one_mode(root, "values")
        _one_mode(root, "formulas")

    print()
    if _failures:
        print(f"FAILED: {len(_failures)} check(s): {_failures}")
        return 1
    print("ALL COMPARISON STYLE-CACHE CHECKS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
