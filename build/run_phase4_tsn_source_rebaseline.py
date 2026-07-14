"""Rebuild every TSN normalized source from one raw-only library, in isolation.

This is a production-integration witness, not the independent semantic oracle.  It
hash-binds the complete registered raw/evidence universe, exercises the public
``tsn_library.build_consolidated`` certification/reuse boundary, and records the exact
normalized workbook shape/result without reading or writing the app's live TSN library.
"""
from __future__ import annotations

import argparse
import contextlib
import hashlib
import importlib
import json
import os
from pathlib import Path
import secrets
import sys
import time
from types import ModuleType


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import artifact_store  # noqa: E402
import consolidation_meta  # noqa: E402
from events import Events  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
import report_catalog  # noqa: E402
import tsn_library  # noqa: E402
import tsn_district_contract as raw_contract  # noqa: E402


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _manifest(files, root):
    """Stable exact manifest retaining the accepted witness's v1 representation.

    ``capture_raw_manifest`` rejects links and any member replaced or mutated while
    read.  The field names here intentionally remain ``bytes``/``sha256`` so the
    already accepted core/evidence aggregate hashes do not change.
    """
    strict, _captured = raw_contract.capture_raw_manifest(list(files), root)
    members = [{
        "relative_path": item["relative_path"],
        "bytes": item["byte_length"],
        "sha256": item["sha256"],
    } for item in strict["members"]]
    return {
        "members": members,
        "member_count": len(members),
        "bytes": sum(item["bytes"] for item in members),
        "sha256": strict["sha256"],
    }


def _inspect_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            it = ws.iter_rows(values_only=True)
            header = list(next(it, ()) or ())
            rows = 0
            routes = set()
            for row in it:
                if not row or not any(value not in (None, "") for value in row):
                    continue
                rows += 1
                if row[0] not in (None, ""):
                    routes.add(str(row[0]))
            sheets.append({
                "name": sheet_name,
                "header": header,
                "data_rows": rows,
                "distinct_first_column_values": len(routes),
            })
        return sheets
    finally:
        wb.close()


def _discover(source_root, specs):
    """Discover the exact registered raw and optional-evidence member universes."""
    core = []
    evidence = []
    by_report = {}
    for spec in specs:
        family_root = source_root / spec.subdir
        raw_dir = family_root / "raw"
        raw_files = sorted(
            (path for path in raw_dir.glob(spec.raw_glob)
             if path.is_file() and not path.name.startswith("~$")),
            key=lambda path: path.relative_to(source_root).as_posix().casefold(),
        )
        if not raw_files:
            raise ValueError(f"{spec.subdir}: no {spec.raw_glob} source in {raw_dir}")
        evidence_files = []
        if spec.evidence_pdfs:
            evidence_dir = family_root / "pdf"
            evidence_files = sorted(
                (path for path in evidence_dir.glob("*.pdf") if path.is_file()),
                key=lambda path: path.relative_to(source_root).as_posix().casefold(),
            )
        by_report[spec.subdir] = {
            "raw": raw_files,
            "evidence": evidence_files,
        }
        core.extend(raw_files)
        evidence.extend(evidence_files)
    return {"core": core, "evidence": evidence, "by_report": by_report}


def _global_family_members(global_manifest, spec):
    prefix = f"{spec.subdir}/raw/"
    result = []
    for member in global_manifest["members"]:
        relative = member["relative_path"]
        if relative.casefold().startswith(prefix.casefold()):
            result.append({
                "relative_path": relative[len(prefix):],
                "byte_length": member["bytes"],
                "sha256": member["sha256"],
            })
    return result


def _manifest_subset(global_manifest, prefix):
    """Return one already-bound subtree in the accepted v1 manifest shape."""
    prefix = prefix.rstrip("/") + "/"
    members = [dict(member) for member in global_manifest["members"]
               if member["relative_path"].casefold().startswith(prefix.casefold())]
    canonical = "".join(
        f"{item['relative_path']}\t{item['bytes']}\t{item['sha256']}\n"
        for item in members).encode("utf-8")
    return {
        "members": members,
        "member_count": len(members),
        "bytes": sum(item["bytes"] for item in members),
        "sha256": hashlib.sha256(canonical).hexdigest(),
    }


def _validate_family_against_global(spec, family_manifest, global_manifest):
    expected = _global_family_members(global_manifest, spec)
    if family_manifest["members"] != expected:
        raise RuntimeError(
            f"{spec.subdir}: raw members/bytes changed after the initial global bind")


def _module_path(module):
    path = Path(module.__file__).resolve()
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError as exc:
        raise ValueError(
            f"operative module {module.__name__!r} is outside the repository: {path}") from exc


def _code_provenance(specs):
    """Bind the runner, shared publication boundary, builders, and projectors."""
    modules = {}

    def add(module, role):
        if not isinstance(module, ModuleType) or not getattr(module, "__file__", None):
            return
        path = Path(module.__file__).resolve()
        try:
            path.relative_to(ROOT)
        except ValueError:
            return
        entry = modules.setdefault(module.__name__, {"module": module, "roles": set()})
        entry["roles"].add(role)

    add(tsn_library, "shared TSN library boundary")
    add(raw_contract, "raw-manifest/district contract")
    add(consolidation_meta, "outcome sidecar boundary")
    add(artifact_store, "atomic workbook publication")
    add(report_catalog, "report/normalizer catalog")
    for spec in specs:
        builder_name = spec.builder.split(":", 1)[0]
        builder_module = importlib.import_module(builder_name)
        add(builder_module, f"{spec.subdir} builder")

        # The statewide shims import their operative compare/projector module as a
        # module global.  Discover those references instead of maintaining a second
        # hand-written registry.  The district sequence comparator has no such
        # builder import, so its catalog-derived conventional name is included too.
        for value in vars(builder_module).values():
            if isinstance(value, ModuleType) and value.__name__.startswith("compare"):
                add(value, f"{spec.subdir} imported compare/projector")
        compare_name = f"compare_{spec.subdir}_tsn"
        if (ROOT / "scripts" / f"{compare_name}.py").is_file():
            add(importlib.import_module(compare_name),
                f"{spec.subdir} conventional compare/projector")

    # Bind repo-local compare/projector dependencies imported by the modules above
    # (notably compare_tsn_common and compare_core) one level deep.
    for entry in list(modules.values()):
        for value in vars(entry["module"]).values():
            if isinstance(value, ModuleType) and value.__name__.startswith("compare"):
                add(value, f"dependency of {entry['module'].__name__}")
                continue
            owner = getattr(value, "__module__", "")
            if isinstance(owner, str) and owner.startswith("compare"):
                try:
                    add(importlib.import_module(owner),
                        f"callable dependency of {entry['module'].__name__}")
                except ImportError:
                    # Only repo-importable dependencies can be operative in this
                    # loaded witness process; a synthetic test callable may name a
                    # non-importable owner and is deliberately ignored.
                    pass

    records = [{
        "relative_path": Path(__file__).resolve().relative_to(ROOT).as_posix(),
        "bytes": Path(__file__).stat().st_size,
        "sha256": _sha256(Path(__file__)),
        "modules": ["__witness_runner__"],
        "roles": ["phase-4 witness runner"],
    }]
    by_path = {}
    for name, entry in modules.items():
        relative = _module_path(entry["module"])
        item = by_path.setdefault(relative, {
            "relative_path": relative,
            "bytes": (ROOT / relative).stat().st_size,
            "sha256": _sha256(ROOT / relative),
            "modules": [],
            "roles": set(),
        })
        item["modules"].append(name)
        item["roles"].update(entry["roles"])
    for item in by_path.values():
        item["modules"].sort(key=str.casefold)
        item["roles"] = sorted(item["roles"], key=str.casefold)
        records.append(item)
    records.sort(key=lambda item: item["relative_path"].casefold())
    canonical = "".join(
        f"{item['relative_path']}\t{item['bytes']}\t{item['sha256']}\n"
        for item in records).encode("utf-8")
    return {
        "members": records,
        "member_count": len(records),
        "bytes": sum(item["bytes"] for item in records),
        "sha256": hashlib.sha256(canonical).hexdigest(),
    }


@contextlib.contextmanager
def _redirect_library(source_root, output_root):
    """Redirect only tsn_library's path functions for this witness attempt."""
    saved = (tsn_library.raw_dir, tsn_library.pdf_dir,
             tsn_library.consolidated_path)

    def redirected_raw(report):
        return source_root / report / "raw"

    def redirected_pdf(report):
        return source_root / report / "pdf"

    def redirected_consolidated(report):
        spec = tsn_library.get(report)
        return output_root / report / "consolidated" / spec.consolidated_name

    tsn_library.raw_dir = redirected_raw
    tsn_library.pdf_dir = redirected_pdf
    tsn_library.consolidated_path = redirected_consolidated
    try:
        yield
    finally:
        (tsn_library.raw_dir, tsn_library.pdf_dir,
         tsn_library.consolidated_path) = saved


def _result_claim(result):
    return {
        "status": getattr(result, "status", None),
        "completion": getattr(result, "completion", None),
        "message": getattr(result, "message", ""),
        "summary_lines": list(getattr(result, "summary_lines", ()) or ()),
        "skipped_inputs": getattr(result, "skipped_inputs", None),
        "failed_inputs": getattr(result, "failed_inputs", None),
    }


def _complete(result):
    return (
        getattr(result, "status", None) == "ok"
        and getattr(result, "completion", None) == "complete"
        and getattr(result, "skipped_inputs", None) == 0
        and getattr(result, "failed_inputs", None) == 0
    )


def _file_identity(path):
    return {
        "bytes": path.stat().st_size,
        "mtime_ns": path.stat().st_mtime_ns,
        "sha256": _sha256(path),
    }


def run(source_root: Path, output_root: Path):
    source_root = source_root.resolve()
    output_root = output_root.resolve()
    if not source_root.is_dir():
        raise ValueError(f"TSN source root does not exist: {source_root}")
    try:
        output_root.relative_to(source_root)
    except ValueError:
        pass
    else:
        raise ValueError("isolated output root must not be inside the raw source root")
    if output_root.exists() and any(output_root.iterdir()):
        raise ValueError(f"isolated output root must be absent or empty: {output_root}")
    output_root.mkdir(parents=True, exist_ok=True)

    specs = tsn_library.reports()
    discovered = _discover(source_root, specs)
    core_manifest = _manifest(discovered["core"], source_root)
    evidence_manifest = _manifest(discovered["evidence"], source_root)
    code_manifest = _code_provenance(specs)
    families = []

    with _redirect_library(source_root, output_root):
        for spec in specs:
            raw_dir = source_root / spec.subdir / "raw"
            family_files = discovered["by_report"][spec.subdir]["raw"]
            pre_builder_manifest = raw_contract.canonical_raw_manifest(
                family_files, raw_dir)
            _validate_family_against_global(spec, pre_builder_manifest, core_manifest)
            family_source_manifest = _manifest_subset(
                core_manifest, f"{spec.subdir}/raw")
            family_evidence_manifest = _manifest_subset(
                evidence_manifest, f"{spec.subdir}/pdf")

            lines = []
            started = time.perf_counter()
            result = tsn_library.build_consolidated(
                spec.subdir,
                events=Events(on_log=lines.append),
                confirm_overwrite=lambda _path: True,
                force=True,
            )
            elapsed = time.perf_counter() - started
            out = tsn_library.consolidated_path(spec.subdir)
            sidecar = consolidation_meta.meta_path(out)
            try:
                builder_manifest = raw_contract.validate_raw_manifest(
                    getattr(result, "tsn_raw_manifest", None))
                certificate_error = None
            except ValueError as exc:
                builder_manifest = None
                certificate_error = str(exc)
            certificate_matches = builder_manifest == pre_builder_manifest
            status_after_build = tsn_library.status(spec.subdir)
            record = {
                "report": spec.subdir,
                "label": spec.label,
                "raw_kind": spec.raw_kind,
                "raw_glob": spec.raw_glob,
                "normalization_version": spec.normalization_version,
                "builder": spec.builder,
                "raw_manifest": family_source_manifest,
                "builder_raw_manifest": builder_manifest,
                "builder_certificate_matches": certificate_matches,
                "builder_certificate_error": certificate_error,
                "optional_evidence_manifest": family_evidence_manifest,
                "elapsed_seconds": round(elapsed, 3),
                "result": _result_claim(result),
                "status_after_build": status_after_build,
                "events": lines,
            }
            if (not _complete(result) or not certificate_matches or not out.is_file()
                    or not sidecar.is_file() or not status_after_build["current"]):
                record["output"] = None
                families.append(record)
                raise RuntimeError(json.dumps(record, indent=2, ensure_ascii=False))

            output_before_reuse = _file_identity(out)
            sidecar_before_reuse = _file_identity(sidecar)
            reuse = tsn_library.build_consolidated(
                spec.subdir,
                events=Events(on_log=lines.append),
                confirm_overwrite=lambda _path: True,
                force=False,
            )
            status_after_reuse = tsn_library.status(spec.subdir)
            output_after_reuse = _file_identity(out)
            sidecar_after_reuse = _file_identity(sidecar)
            reuse_certified = (
                _complete(reuse)
                and "reused" in getattr(reuse, "message", "").casefold()
                and status_after_reuse["current"]
                and output_after_reuse == output_before_reuse
                and sidecar_after_reuse == sidecar_before_reuse
            )
            record["reuse"] = {
                "result": _result_claim(reuse),
                "status": status_after_reuse,
                "output_unchanged": output_after_reuse == output_before_reuse,
                "sidecar_unchanged": sidecar_after_reuse == sidecar_before_reuse,
                "certified": reuse_certified,
            }
            if not reuse_certified:
                record["output"] = None
                families.append(record)
                raise RuntimeError(json.dumps(record, indent=2, ensure_ascii=False))
            record["output"] = {
                "relative_path": out.relative_to(output_root).as_posix(),
                "bytes": out.stat().st_size,
                "sha256": _sha256(out),
                "sidecar_relative_path": sidecar.relative_to(output_root).as_posix(),
                "sidecar_sha256": _sha256(sidecar),
                "sheets": _inspect_xlsx(out),
            }
            families.append(record)

    # Re-discover, re-read, and bind every registered core/evidence member after
    # the last family.  This catches late addition/removal/replacement, including
    # evidence PDFs no normalizer itself opens.
    final_discovered = _discover(source_root, specs)
    final_core = _manifest(final_discovered["core"], source_root)
    final_evidence = _manifest(final_discovered["evidence"], source_root)
    if final_core != core_manifest or final_evidence != evidence_manifest:
        raise RuntimeError("registered TSN raw/evidence universe changed during witness")
    final_code = _code_provenance(specs)
    if final_code != code_manifest:
        raise RuntimeError("operative witness code changed during witness")

    expected_output_members = set()
    for family in families:
        expected_output_members.add(family["output"]["relative_path"])
        expected_output_members.add(family["output"]["sidecar_relative_path"])
    actual_output_files = sorted(
        (path for path in output_root.rglob("*") if path.is_file()),
        key=lambda path: path.relative_to(output_root).as_posix().casefold(),
    )
    actual_output_members = {
        path.relative_to(output_root).as_posix() for path in actual_output_files}
    if actual_output_members != expected_output_members:
        extra = sorted(actual_output_members - expected_output_members, key=str.casefold)
        missing = sorted(expected_output_members - actual_output_members, key=str.casefold)
        raise RuntimeError(
            f"isolated output universe is not exact (extra={extra}; missing={missing})")
    output_manifest = _manifest(actual_output_files, output_root)

    return {
        "schema_version": 2,
        "acceptance": "complete",
        "source_root": str(source_root),
        "output_root": str(output_root),
        "core_raw_manifest": core_manifest,
        "core_raw_manifest_after": final_core,
        "optional_evidence_manifest": evidence_manifest,
        "optional_evidence_manifest_after": final_evidence,
        "source_universe_stable": True,
        "code_provenance_manifest": code_manifest,
        "code_provenance_manifest_after": final_code,
        "code_provenance_stable": True,
        # This is the exact generated workbook+sidecar universe before the
        # acceptance JSON is published.  ``--result`` may intentionally live
        # under output_root, but it is witness metadata, not a generated TSN
        # normalization artifact.
        "generated_output_artifact_manifest": output_manifest,
        "generated_output_artifact_universe_exact": True,
        "expected_family_count": len(specs),
        "completed_family_count": len(families),
        "families": families,
    }


def _validate_accepted(payload):
    families = payload.get("families")
    expected = payload.get("expected_family_count")
    if (payload.get("acceptance") != "complete"
            or not payload.get("source_universe_stable")
            or not payload.get("code_provenance_stable")
            or not payload.get("generated_output_artifact_universe_exact")
            or not isinstance(families, list)
            or payload.get("completed_family_count") != expected
            or len(families) != expected
            or any(not family.get("reuse", {}).get("certified")
                   or not family.get("builder_certificate_matches")
                   or family.get("output") is None for family in families)):
        raise ValueError("refusing to publish a partial or uncertified witness result")


def write_result_atomic(path, payload):
    """Atomically accept only a fully certified witness payload."""
    _validate_accepted(payload)
    path = Path(path).resolve()
    source_root = Path(payload["source_root"]).resolve()
    try:
        path.relative_to(source_root)
    except ValueError:
        pass
    else:
        raise ValueError("result JSON must not be written inside the raw source root")
    output_root = Path(payload["output_root"]).resolve()
    artifact_paths = set()
    for family in payload["families"]:
        output = family["output"]
        artifact_paths.add((output_root / output["relative_path"]).resolve())
        artifact_paths.add((output_root / output["sidecar_relative_path"]).resolve())
    if path in artifact_paths:
        raise ValueError("result JSON path collides with a witnessed output artifact")
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.tmp-{os.getpid()}-{secrets.token_hex(8)}")
    try:
        with temp.open("x", encoding="utf-8", newline="\n") as stream:
            json.dump(payload, stream, indent=2, sort_keys=True, ensure_ascii=False)
            stream.write("\n")
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temp, path)
    finally:
        try:
            temp.unlink()
        except FileNotFoundError:
            pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", required=True, type=Path)
    parser.add_argument("--output-root", required=True, type=Path)
    parser.add_argument("--result", required=True, type=Path)
    args = parser.parse_args()
    if args.result.resolve().exists():
        raise ValueError(
            f"result path already exists; use a fresh attempt path: {args.result.resolve()}")
    payload = run(args.source_root, args.output_root)
    write_result_atomic(args.result, payload)
    print(f"TSN raw manifest: {payload['core_raw_manifest']['member_count']} members / "
          f"{payload['core_raw_manifest']['bytes']} bytes / "
          f"{payload['core_raw_manifest']['sha256']}")
    for family in payload["families"]:
        sheet = family["output"]["sheets"][0]
        print(f"{family['report']}: {family['result']['status']} / "
              f"{sheet['data_rows']} rows / {family['output']['sha256']}")


if __name__ == "__main__":
    main()
