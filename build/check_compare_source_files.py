"""Golden checks for the "Source Files" provenance companion sheet
(scripts/compare_tsn_common.source_files_from_consolidated +
.write_source_files_sheet).

Every comparison workbook gets a default "Source Files" sheet naming which per-route
export each TSMIS row came from — `<report-prefix>_route_<route>.<ext>`. The route is
the CONSOLIDATED workbook's leading Route column (the FILE route), so it names the
real source even when the compared route differs (an Intersection Detail equate). The
sheet is written write-only-safe (create_sheet + append), so it needs no change to
the correctness-locked engine.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_source_files.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_tsn_common as ctc
from openpyxl import Workbook, load_workbook


def test_source_files_from_consolidated():
    """Filenames come from column 0, filtered/ordered exactly like the loaders
    (row_has_data), and suffixed routes + a custom extension are honored."""
    d = Path(tempfile.mkdtemp())
    p = d / "cons.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Route", "A", "B"])
    ws.append(["001", "x", "y"])
    ws.append([None, None, None])            # a blank row: filtered out
    ws.append(["005S", "p", "q"])
    wb.save(p)
    got = ctc.source_files_from_consolidated(str(p), "Data", "intersection_detail")
    assert got == ["intersection_detail_route_001.xlsx",
                   "intersection_detail_route_005S.xlsx"], got
    # a PDF-sourced flavor names the .pdf
    got_pdf = ctc.source_files_from_consolidated(str(p), "Data", "hl", ext="pdf")
    assert got_pdf == ["hl_route_001.pdf", "hl_route_005S.pdf"], got_pdf
    # a missing sheet / unreadable path is a logged no-op, never a crash
    assert ctc.source_files_from_consolidated(str(p), "Nope", "x") == []
    assert ctc.source_files_from_consolidated(str(d / "gone.xlsx"), "Data", "x") == []
    # a bare per-route workbook (col 0 is NOT a leading 'Route') is skipped — we
    # never name a data cell as a route
    pr = d / "perroute.xlsx"
    w2 = Workbook()
    s2 = w2.active
    s2.title = "Data"
    s2.append(["County", "PM", "Desc"])
    s2.append(["ORA", "001.0", "x"])
    w2.save(pr)
    assert ctc.source_files_from_consolidated(str(pr), "Data", "highway_log") == []


def test_source_files_from_rows():
    """The cross-env / baseline path names each row from the route prepended at
    column 0 (no consolidated workbook to read)."""
    rows = [["001", "a"], ["005S", "b"], [None, "c"]]
    assert ctc.source_files_from_rows(rows, "tsar_ramp_detail", "xlsx") == [
        "tsar_ramp_detail_route_001.xlsx",
        "tsar_ramp_detail_route_005S.xlsx", ""], rows
    assert ctc.source_files_from_rows([["014U"]], "intersection_detail", "pdf") == [
        "intersection_detail_route_014U.pdf"]


def test_write_source_files_sheet_write_only():
    """The companion sheet is written into a WRITE-ONLY workbook (the streaming
    shape the real comparison uses) and lists side/row/route/file per row."""
    d = Path(tempfile.mkdtemp())
    p = d / "cmp.xlsx"
    wb = Workbook(write_only=True)
    wb.create_sheet("Comparison").append(["placeholder"])   # a prior streamed sheet
    rows = [["001", "a"], ["005S", "b"]]
    files = ["intersection_detail_route_001.xlsx",
             "intersection_detail_route_005S.xlsx"]
    ctc.write_source_files_sheet(wb, [("TSMIS", rows, files)])
    wb.save(p)
    rb = load_workbook(p, read_only=True, data_only=True)
    assert "Source Files" in rb.sheetnames, rb.sheetnames
    it = rb["Source Files"].iter_rows(values_only=True)
    assert list(next(it)) == ["Side", "Row #", "Route (as compared)", "Source File"]
    body = [list(r) for r in it]
    rb.close()
    assert body == [["TSMIS", 1, "001", "intersection_detail_route_001.xlsx"],
                    ["TSMIS", 2, "005S", "intersection_detail_route_005S.xlsx"]], body


def main():
    test_source_files_from_consolidated()
    test_source_files_from_rows()
    test_write_source_files_sheet_write_only()
    print("OK  Source Files provenance: filenames come from the consolidated Route "
          "column (the file route), blank rows are filtered, PDF flavors name the "
          ".pdf, unreadable inputs log-and-skip, and the companion sheet writes "
          "into the write-only comparison workbook.")


if __name__ == "__main__":
    main()
