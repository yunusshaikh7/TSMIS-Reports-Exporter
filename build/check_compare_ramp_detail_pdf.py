"""CMP-AUD-036 — the Ramp Detail (PDF) source gate requires the full print shape.

`compare_ramp_detail_pdf._load_tsmis_pdf` used to accept any workbook with `PM`
among the first five header cells and `On/Off` anywhere, then expand each row by
position. A truncated four-column `Route/Location/PM/On-Off` workbook was accepted
and every absent field fabricated as blank; an Excel-consolidated pick (no
print-only columns) could also slip through. The gate now requires the EXACT
PDF-consolidated width and the two trailing print-only sentinels (On/Off, Ramp
Type) in order. Both PDF-vs-TSN and PDF-vs-Excel ride this loader.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_ramp_detail_pdf.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_ramp_detail_pdf as rdp
import compare_ramp_detail_tsn as _rd
import consolidate_tsmis_ramp_detail_pdf as cons
from openpyxl import Workbook

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _valid_header():
    # The real PDF-consolidated header: ["Route"] + the consolidator's print
    # HEADER, with the shifted/blank labels rendered as empty strings.
    return ["Route"] + [c if c is not None else "" for c in cons.HEADER]


def _write(path, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = _rd.TSMIS_SHEET
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _refused(path):
    try:
        rdp._load_tsmis_pdf(str(path))
        return None
    except ValueError as e:
        return str(e)


def test_width_mirror():
    print("the gate's expected width mirrors the consolidator (CMP-AUD-036):")
    check("_PDF_WIDTH == 1 + len(consolidator HEADER)",
          rdp._PDF_WIDTH == 1 + len(cons.HEADER))
    check("the print-only sentinels are the HEADER's last two labels",
          rdp._PDF_SENTINELS == tuple(cons.HEADER[-2:]) == ("On/Off", "Ramp Type"))


def test_gate():
    print("the PDF source gate requires the full print shape:")
    with tempfile.TemporaryDirectory() as d:
        d = Path(d)
        valid_h = _valid_header()
        # A full PDF-consolidated workbook loads. The row is read BY POSITION
        # (Route0 Location1 PR2 PM3 Date4 sfx5 HG6 Area4-7 City8 R/U9 Desc10
        # blank11 On/Off12 RampType13), so give Location a real "01-DN-101".
        valid_row = ["001", "01-DN-101", "R", "1.000", "2026-01-01", "", "D",
                     "Y", "C", "U", "9/DESC", "", "N", "D"]
        good = d / "good.xlsx"
        _write(good, valid_h, [valid_row])
        try:
            rows, has_route = rdp._load_tsmis_pdf(str(good))
            check("a full PDF-consolidated workbook is accepted",
                  has_route is True and len(rows) == 1)
        except ValueError as e:
            check("a full PDF-consolidated workbook is accepted", False)
            print("     ->", e)

        # The finding's exact fabricated four-column shape refuses.
        trunc4 = d / "trunc4.xlsx"
        _write(trunc4, ["Route", "Location", "PM", "On/Off"],
               [["001", "01-DN-101", "1.000", "N"]])
        check("a truncated four-column Route/Location/PM/On-Off workbook refuses",
              "PDF-CONSOLIDATED" in (_refused(trunc4) or ""))

        # Both sentinels present but the row is still truncated (< full width).
        trunc5 = d / "trunc5.xlsx"
        _write(trunc5, ["Route", "Location", "PM", "On/Off", "Ramp Type"],
               [["001", "01-DN-101", "1.000", "N", "D"]])
        check("a truncated workbook carrying both sentinels still refuses",
              _refused(trunc5) is not None)

        # An Excel-consolidated pick (the print-only columns dropped) refuses.
        excel = d / "excel.xlsx"
        _write(excel, ["Route"] + [c if c is not None else "" for c in cons.HEADER[:-2]],
               [["001"] + ["x"] * (len(cons.HEADER) - 2)])
        check("an Excel-consolidated pick (no print-only columns) refuses",
              _refused(excel) is not None)

        # Every prefix truncation of the valid header refuses; the full one is
        # the only accepted width.
        all_prefixes_refuse = all(
            _refused_prefix(d, valid_h, k) for k in range(1, len(valid_h)))
        check("every prefix truncation of the valid header refuses",
              all_prefixes_refuse)


def _refused_prefix(d, valid_h, k):
    p = d / f"pre{k}.xlsx"
    _write(p, valid_h[:k], [["001"] + ["x"] * (k - 1)])
    return _refused(p) is not None


def test_side_labels():
    """CMP-AUD-069: the missing-input existence message uses the flavor's OWN
    side labels, not the shared driver's TSMIS/TSN defaults — so a PDF-vs-Excel
    run with a missing second file says 'TSMIS (Excel)', never 'TSN'."""
    print("missing-input diagnostics carry the flavor's side labels (CMP-AUD-069):")
    from events import Events
    with tempfile.TemporaryDirectory() as d:
        d = Path(d)
        exists = d / "exists.xlsx"
        exists.write_bytes(b"x")            # existence check only stats the path
        missing = d / "missing.xlsx"
        out = d / "out.xlsx"

        def msg(flavor, a, b):
            r = flavor.compare(str(a), str(b), str(out), events=Events(),
                               confirm_overwrite=lambda _p: True)
            return r.status, (r.message or "")

        st, m = msg(rdp.TSMIS_PDF_VS_EXCEL, exists, missing)
        check("PDF-vs-Excel missing 2nd file names 'TSMIS (Excel)', not 'TSN'",
              st == "error" and "TSMIS (Excel)" in m and "TSN" not in m)
        st, m = msg(rdp.TSMIS_PDF_VS_TSN, exists, missing)
        check("PDF-vs-TSN missing 2nd file names 'TSN'", st == "error" and "TSN" in m)
        st, m = msg(rdp.TSMIS_PDF_VS_EXCEL, missing, exists)
        check("missing 1st file names 'TSMIS (PDF)' (not the default 'TSMIS')",
              st == "error" and "TSMIS (PDF)" in m)


def test_self_check_null_parity():
    """HF-04 / PCOA-FINAL-012: the same-source PDF-vs-Excel self check over
    rows carrying the print's null tokens ('-' in Area 4 / OF, 'NO RAMP LINEAR
    EVENT' in Description) and print On/Off letters on BOTH sides reports ZERO
    differing cells — the PDF leg has always projected those render artifacts,
    and the July-2026 Excel export now carries the same tokens, so the Excel
    leg must project them symmetrically. A classic-layout Excel side (blank
    cells) stays zero too."""
    print("same-source null-token symmetry (PCOA-FINAL-012):")
    from events import Events
    from openpyxl import load_workbook

    hdr_2026 = ["Route", "Location", "PRE", "PM", "Date of Record", "HG",
                "Area 4", "City Code", "R/U", "OF", "TY", "Description"]
    with tempfile.TemporaryDirectory() as d:
        d = Path(d)
        # The PDF-consolidated side: one null-token ramp (the 005 / 025.218
        # class) + one ordinary on-ramp (print letter N).
        pdf_path = d / "pdf.xlsx"
        _write(pdf_path, _valid_header(), [
            ["005", "07-LA-005", "", "25.218", "01/01/2014", "", "D", "-",
             "LA", "U", "NO RAMP LINEAR EVENT", "", "-", "-"],
            ["005", "07-LA-005", "R", "26.000", "01/01/2014", "", "D", "Y",
             "LA", "U", "005/NB ON FR X", "", "N", "C"],
        ])
        # The July-2026 Excel side of the SAME report: the same two rows as
        # the new export renders them ('-' / NRLE / the print letter N).
        excel_new = d / "excel_new.xlsx"
        _write(excel_new, hdr_2026, [
            ["005", "07-LA-005", "", "25.218", "01/01/2014", "D", "-", "LA",
             "U", "-", "-", "NO RAMP LINEAR EVENT"],
            ["005", "07-LA-005", "R", "26.000", "01/01/2014", "D", "Y", "LA",
             "U", "N", "C", "005/NB ON FR X"],
        ])

        def run(excel_path, tag):
            out = d / f"cmp_{tag}.xlsx"
            res = rdp.TSMIS_PDF_VS_EXCEL.compare(
                str(pdf_path), str(excel_path), str(out), events=Events(),
                confirm_overwrite=lambda _p: True, mode="values")
            return res, out

        res, out = run(excel_new, "new")
        # A pre-fix tree refuses the July-2026 Excel side outright
        # (PCOA-FINAL-001), so degrade the remaining assertions to clean FAILs
        # instead of crashing on the never-written workbook.
        ran = res.status == "ok"
        check("the self check runs over the July-2026 Excel side", ran)
        counts = res.comparison_outcome.counts
        check("ZERO differing cells across a fully-paired same-source pair "
              "(pre-fix: refused entirely; post-012-regression: 3 per "
              "null-token row)",
              ran and counts.known and counts.differing_cells == 0)
        check("both rows pair (no one-sided rows)",
              ran and counts.paired_rows == 2 and counts.side_a_only_rows == 0
              and counts.side_b_only_rows == 0)
        check("the verdict is a clean match", res.verdict == "match")
        # The workbook itself: no ≠ marker anywhere, and the On/Off context
        # column reads the SAME projected letter from both legs.
        if ran and out.exists():
            wb = load_workbook(out, read_only=True, data_only=True)
            body = list(wb["Comparison"].iter_rows(values_only=True))
            wb.close()
            header = [("" if c is None else str(c)) for c in body[0]]
            neq = sum(1 for r in body[1:] for v in r
                      if isinstance(v, str) and " ≠ " in v)
            check("no ≠ marker in the Comparison sheet", neq == 0)
            onoff = [str(r[header.index("On/Off")]) for r in body[1:]]
            check("On/Off context shows the projected O (print N -> O, both legs)",
                  "O" in onoff and "N" not in onoff and "-" not in onoff)
        else:
            check("no ≠ marker in the Comparison sheet", False)
            check("On/Off context shows the projected O (print N -> O, both legs)",
                  False)

        # Regression: a classic-layout Excel side (blank where the print marks
        # '-') still reports zero — the projection is a no-op on blanks.
        # Classic row positions: Route, Location, PR, PM, Date, sfx, HG,
        # Area4, City, R/U, Desc, blank — the null cells are BLANK here.
        excel_old = d / "excel_old.xlsx"
        wbk = Workbook()
        ws = wbk.active
        ws.title = _rd.TSMIS_SHEET
        ws.append(["Route"] + list(_rd._TSMIS_HEADER[1:]))
        ws.append(["005", "07-LA-005", "", "25.218", "01/01/2014", "", "D",
                   "", "LA", "U", "", ""])
        ws.append(["005", "07-LA-005", "R", "26.000", "01/01/2014", "", "D",
                   "Y", "LA", "U", "005/NB ON FR X", ""])
        wbk.save(excel_old)
        wbk.close()
        res_old, _ = run(excel_old, "old")
        check("the classic Excel side still reports zero differing cells",
              res_old.status == "ok"
              and res_old.comparison_outcome.counts.differing_cells == 0)


def main():
    test_width_mirror()
    test_gate()
    test_side_labels()
    test_self_check_null_parity()
    print()
    if _fail:
        print(f"{len(_fail)} CHECK(S) FAILED:")
        for f in _fail:
            print("  -", f)
        sys.exit(1)
    print("ALL COMPARE-RAMP-DETAIL-PDF CHECKS PASSED")


if __name__ == "__main__":
    main()
