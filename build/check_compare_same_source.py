"""CMP-AUD-067: same-source (PDF-vs-Excel) comparisons must not reuse
cross-system projections that erase the render differences they exist to
detect.

The vs-TSN projectors deliberately reconcile unlike TSMIS/TSN encodings
(control-type crosswalks, NA folds, canonical-key inference). Reusing them in
the SAME-SOURCE flavors made real PDF/Excel render differences invisible —
the finding's isolated mutations returned "EVERYTHING MATCHES". This check IS
that mutation matrix, run per family through the real flavor compare()
(fixtures carry the CMP-AUD-066 PDF marker on the PDF side):

  * Highway Sequence — FIXED by CMP-AUD-199/204 (its own same-source loader);
    pinned here so it stays fixed: a lost "001/" Description prefix FLAGS.
  * Ramp Detail — never had an instance (its loader is verbatim); pinned:
    a Description mutation flags verbatim.
  * Intersection Detail — the "Ctrl Type" J→S crosswalk must NOT apply
    between two TSMIS renders: PDF `J` vs Excel `S` FLAGS, and a J-vs-A
    difference displays the RAW "J ≠ A" (not the rewritten "S ≠ A").
  * Highway Detail joins when its same-source projection lands (pm_canon
    HG-fill + NA crosswalk — tracked in the 067 plan census).

CI-safe: pure Python fixtures, no local data.
"""
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import Workbook, load_workbook  # noqa: E402

import compare_highway_detail_pdf as hdp  # noqa: E402
import compare_highway_sequence_pdf as hslp  # noqa: E402
import compare_intersection_detail_pdf as idp  # noqa: E402
import compare_ramp_detail_pdf as rdp  # noqa: E402
import consolidate_tsmis_ramp_detail_pdf as rdc  # noqa: E402
import highway_detail_columns as hdc  # noqa: E402
import intersection_detail_columns as idc  # noqa: E402
from events import Events  # noqa: E402
from pdf_table_lib import write_pdf_source_marker  # noqa: E402

failures = []


def check(label, cond, detail=""):
    print(("OK   " if cond else "FAIL ") + label
          + (f" — {detail}" if detail and not cond else ""))
    if not cond:
        failures.append(label)


def write_wb(path, sheet, header, rows, marked=False):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    if marked:
        write_pdf_source_marker(wb)
    wb.save(path)
    wb.close()


def run(flavor, a, b, out):
    return flavor.compare(str(a), str(b), str(out), events=Events(),
                          confirm_overwrite=lambda _p: True, mode="values")


def diff_cells(path, limit=8):
    wb = load_workbook(path)
    try:
        return [v for row in wb["Comparison"].iter_rows(values_only=True)
                for v in row if isinstance(v, str) and " ≠ " in v][:limit]
    finally:
        wb.close()


def summary_says_diff(res):
    return (res.status == "ok"
            and any("DIFFERENCES FOUND" in ln for ln in res.summary_lines))


def summary_says_match(res):
    return (res.status == "ok"
            and any("EVERYTHING MATCHES" in ln for ln in res.summary_lines))


tmp = Path(tempfile.mkdtemp(prefix="tsmis_same_source_"))

# --------------------------------------------------------------------------- #
# Highway Sequence — the fixed state stays fixed
# --------------------------------------------------------------------------- #
print("Highway Sequence (fixed by CMP-AUD-199/204 — pinned):")
HSL_HEADER = ["Route", "County", "City", "PR", "PM", "SFX", "HG", "FT",
              "Distance", "Description"]
a = tmp / "hsl_pdf.xlsx"
b = tmp / "hsl_xls.xlsx"
write_wb(a, "Highway Locations", HSL_HEADER,
         [["001", "ALP", "", "", "001.000", "", "D", "H", "", "001/JCT 5"]],
         marked=True)
write_wb(b, "Highway Locations", HSL_HEADER,
         [["001", "ALP", "", "", "001.000", "", "D", "H", "", "JCT 5"]])
res = run(hslp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "hsl_cmp.xlsx")
check("HSL: a lost own-route Description prefix FLAGS (verbatim both sides)",
      summary_says_diff(res),
      f"{res.status}: {(res.summary_lines or [res.message])[0][:160]}")

# --------------------------------------------------------------------------- #
# Ramp Detail — verbatim loader, pinned
# --------------------------------------------------------------------------- #
print("Ramp Detail (verbatim — pinned):")
RD_HEADER = ["Route"] + list(rdc.HEADER)


def rd_row(desc):
    r = [None] * len(RD_HEADER)
    r[0], r[1], r[3], r[4] = "001", "04-CC-001", "000.198", "10/01/1996"
    r[10] = desc
    return r


a = tmp / "rd_pdf.xlsx"
b = tmp / "rd_xls.xlsx"
write_wb(a, rdc.SHEET_NAME, RD_HEADER, [rd_row("EB ON FR MAIN")], marked=True)
write_wb(b, rdc.SHEET_NAME, RD_HEADER, [rd_row("EB ON FR OTHER")])
res = run(rdp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "rd_cmp.xlsx")
check("RD: a Description mutation flags verbatim", summary_says_diff(res),
      f"{res.status}: {(res.summary_lines or [res.message])[0][:160]}")

# --------------------------------------------------------------------------- #
# Intersection Detail — the J→S crosswalk must not apply between two renders
# --------------------------------------------------------------------------- #
print("Intersection Detail (CMP-AUD-067):")
ID_HEADER = ["Route"] + list(idc.HEADER)
CTRL_TYPE_I = ID_HEADER.index("Ctrl Type")      # the folded field (probe-proven)


def id_row(ct):
    r = [None] * len(ID_HEADER)
    r[0] = "001"
    r[1], r[2], r[4], r[5] = "", "000.100", "04 SOL 001", "73-10-19"
    r[CTRL_TYPE_I] = ct
    r[21] = "MAIN ST"
    return r


a = tmp / "id_pdf.xlsx"
b = tmp / "id_xls.xlsx"
write_wb(a, "Intersection Detail", ID_HEADER, [id_row("J")], marked=True)
write_wb(b, "Intersection Detail", ID_HEADER, [id_row("S")])
res = run(idp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "id_cmp1.xlsx")
check("ID: PDF Ctrl Type J vs Excel S FLAGS (no cross-system fold between renders)",
      summary_says_diff(res),
      f"{res.status}: {(res.summary_lines or [res.message])[0][:160]}")

write_wb(a, "Intersection Detail", ID_HEADER, [id_row("J")], marked=True)
write_wb(b, "Intersection Detail", ID_HEADER, [id_row("A")])
res = run(idp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "id_cmp2.xlsx")
cells = diff_cells(tmp / "id_cmp2.xlsx") if res.status == "ok" else []
check("ID: a J-vs-A difference displays the RAW values (J ≠ A, not S ≠ A)",
      summary_says_diff(res) and any(c == "J ≠ A" for c in cells)
      and not any(c == "S ≠ A" for c in cells), str(cells))

write_wb(a, "Intersection Detail", ID_HEADER, [id_row("S")], marked=True)
write_wb(b, "Intersection Detail", ID_HEADER, [id_row("S")])
res = run(idp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "id_cmp3.xlsx")
check("ID: identical renders still MATCH", summary_says_match(res),
      f"{res.status}: {(res.summary_lines or [res.message])[0][:160]}")

# --------------------------------------------------------------------------- #
# Highway Detail — canonical pairing stays, but the raw Post Mile must
# SURFACE (the vs-TSN pm_canon HG-fill made the lost R/L invisible — the
# projected token was also the key) and the TSN-only NA crosswalk must not
# apply between two TSMIS renders.
# --------------------------------------------------------------------------- #
print("Highway Detail (CMP-AUD-067):")
HD_HEADER = ["Route"] + list(hdc.HEADER)
HD_HG_I = HD_HEADER.index("HG")
HD_NA_I = HD_HEADER.index("NA")


def hd_row(pm, hg="R", na=None):
    r = [None] * len(HD_HEADER)
    r[0], r[1], r[2], r[3] = "001", pm, "0.100", "73-10-19"
    r[HD_HG_I] = hg
    if na is not None:
        r[HD_NA_I] = na
    return r


a = tmp / "hd_pdf.xlsx"
b = tmp / "hd_xls.xlsx"
write_wb(a, "Highway Detail", HD_HEADER, [hd_row("000.100R")], marked=True)
write_wb(b, "Highway Detail", HD_HEADER, [hd_row("000.100")])
res = run(hdp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "hd_cmp1.xlsx")
cells = diff_cells(tmp / "hd_cmp1.xlsx") if res.status == "ok" else []
check("HD: a dropped roadbed letter PAIRS canonically but SURFACES as a raw-PM cell",
      summary_says_diff(res) and any("000.100R" in c and "000.100" in c
                                     for c in cells), f"{res.status}: {cells}")

write_wb(a, "Highway Detail", HD_HEADER, [hd_row("000.100R", na="")], marked=True)
write_wb(b, "Highway Detail", HD_HEADER, [hd_row("000.100R", na="A")])
res = run(hdp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "hd_cmp2.xlsx")
check("HD: NA blank-vs-A FLAGS (no TSN crosswalk between renders)",
      summary_says_diff(res),
      f"{res.status}: {(res.summary_lines or [res.message])[0][:160]}")

write_wb(a, "Highway Detail", HD_HEADER, [hd_row("000.100R", na="N")], marked=True)
write_wb(b, "Highway Detail", HD_HEADER, [hd_row("000.100R", na="N")])
res = run(hdp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "hd_cmp3.xlsx")
check("HD: identical renders still MATCH", summary_says_match(res),
      f"{res.status}: {(res.summary_lines or [res.message])[0][:160]}")

# --------------------------------------------------------------------------- #
# Highway Log — the roadbed-canonical key keeps PAIRING the vendor Excel's
# dropped roadbed letters (correctness-locked §7b semantics, unchanged), but
# the raw printed Location must SURFACE as its own compared cell in the
# same-source flavor: an Excel "1.000" whose dittoed Left block implies R used
# to match the PDF's explicit "1.000R" with zero differences.
# --------------------------------------------------------------------------- #
print("Highway Log (CMP-AUD-067):")
import highway_log_columns as hlc  # noqa: E402
import compare_highway_log_pdf as hlp  # noqa: E402


def hl_row(loc, ditto_left=False):
    r = [None] * 31
    r[0], r[1], r[2] = loc, "R1", "DESC"
    if ditto_left:
        for i in hlc.LEFT_BLOCK_IDX:
            r[i] = "+"
    return r


def write_hl(path, rows, marked=False):
    write_wb(path, "Highway Log", list(hlc.HEADER), rows, marked=marked)


a = tmp / "hl pdf route 9.xlsx"
b = tmp / "hl excel route 9.xlsx"
write_hl(a, [hl_row("1.000R")], marked=True)
write_hl(b, [hl_row("1.000", ditto_left=True)])
res = run(hlp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "hl_cmp1.xlsx")
cells = diff_cells(tmp / "hl_cmp1.xlsx") if res.status == "ok" else []
check("HL: a dropped roadbed letter PAIRS canonically but SURFACES as a raw cell",
      summary_says_diff(res) and any("1.000R" in c and "1.000" in c
                                     for c in cells), f"{res.status}: {cells}")

write_hl(a, [hl_row("1.000R")], marked=True)
write_hl(b, [hl_row("1.000R", ditto_left=True)])
res = run(hlp.TSMIS_PDF_VS_EXCEL, a, b, tmp / "hl_cmp2.xlsx")
check("HL: identical raw Locations still MATCH (dittos stay non-asserting)",
      summary_says_match(res),
      f"{res.status}: {(res.summary_lines or [res.message])[0][:160]}")

import shutil  # noqa: E402
shutil.rmtree(tmp, ignore_errors=True)

if failures:
    print(f"\nFAILED: {len(failures)}")
    sys.exit(1)
print("\nSame-source projection matrix (CMP-AUD-067): PASS")
