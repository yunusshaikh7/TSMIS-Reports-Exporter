"""Golden check for COMPARE-KEY-IS-FIRST-COLUMN (scripts/compare_core.py +
compare_env.py).

Reproduces the coarse-key misalignment that inflates the cross-environment
Highway Sequence comparison: the engine used to hard-key every row on the
report's FIRST column. When that column is coarse (Highway Sequence inherits
County, which repeats for hundreds of rows), rows align POSITIONALLY within the
coarse group — so one extra row on one side cascades into spurious field diffs
and mis-counted one-sided rows. The granular column (postmile / PM) is the real
identity; CompareSchema.key_field now selects it.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_keyfield.py

Excel-COM-verified on the real PROD-vs-DEV Highway Sequence subset (2026-06-16):
    key on County (old): 15,797 differing cells / 181 one-sided  (misaligned)
    key on PM   (new):    5,056 differing cells / 221 one-sided, SELF-CHECK 9/9 OK
    Highway Log (key_field 0, unchanged): cell-for-cell IDENTICAL before/after.

This script guards the fix WITHOUT Excel by asserting (1) the pure-Python diff
counts collapse when keyed on the granular column, and (2) the live-formulas
workbook puts the chosen key column in the identity slot while keeping every
other column (in display order) as a field — and that key_field == 0 reproduces
the original [1..n] field order exactly (the regression-locked default path).
"""
import os
import sys
import tempfile
from dataclasses import replace

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from compare_core import (CompareSchema, count_diffs, keys_for, run_compare,
                          union_keys)
from openpyxl import load_workbook

# A coarse first column (County, identical on every row) + a granular key (PM).
# Side B inserts one extra postmile in the middle. Positional (County) keying
# mis-pairs every row after the insert; PM keying isolates the one new row.
HDR = ["County", "PM", "Desc"]
ROWS_A = [["X", "1.0", "a"], ["X", "2.0", "b"],
          ["X", "3.0", "c"], ["X", "4.0", "d"]]
ROWS_B = [["X", "1.0", "a"], ["X", "2.0", "b"], ["X", "2.5", "NEW"],
          ["X", "3.0", "c"], ["X", "4.0", "d"]]


def _counts(sc, key_field):
    kt = keys_for(ROWS_A, False, key_field)
    kn = keys_for(ROWS_B, False, key_field)
    u = union_keys(kt, kn)
    return count_diffs(sc, ROWS_A, ROWS_B, kt, kn, u, False)


def main():
    sc_county = CompareSchema(report_name="KF", header=HDR, key_field=0,
                              side_a="AENV", side_b="BENV", id_noun="row",
                              id_noun_plural="rows", sides_noun="environments")
    sc_pm = replace(sc_county, key_field=1)

    # field_indices: the displayed (non-key) columns, in order.
    assert sc_county.field_indices == [1, 2], sc_county.field_indices
    assert sc_pm.field_indices == [0, 2], sc_pm.field_indices

    # 1) The bug: keyed on County, the single inserted row cascades into spurious
    #    field diffs (rows 3.0/4.0 mis-pair) — both still 4 by coarse tuple.
    c0 = _counts(sc_county, 0)
    assert c0["both"] == 4 and c0["n_only"] == 1, c0
    assert c0["diff_cells"] == 4 and c0["diff_rows"] == 2, \
        ("coarse County key should mis-pair into spurious diffs", c0)

    # 2) The fix: keyed on PM, the matched postmiles are identical and the one
    #    new postmile is correctly B-only — ZERO spurious diffs.
    c1 = _counts(sc_pm, 1)
    assert c1["both"] == 4 and c1["t_only"] == 0 and c1["n_only"] == 1, c1
    assert c1["diff_cells"] == 0 and c1["diff_rows"] == 0, \
        ("granular PM key should eliminate the cascade", c1)

    # 3) Structural: the live workbook puts the key column in the identity slot
    #    and keeps the rest as fields in display order.
    out = os.path.join(tempfile.gettempdir(), "_keyfield_check.xlsx")
    res = run_compare(sc_pm, ROWS_A, ROWS_B, False, out, mode="formulas")
    assert res.status == "ok", res.status
    wb = load_workbook(out, data_only=False)
    header_cells = list(next(wb["Comparison"].iter_rows(max_row=1)))
    # E1 appends versioned hidden state-mask columns after every visible field,
    # and CMP-AUD-218 appends the hidden row-token column after those; key-field
    # geometry is the user-visible prefix, not those internal twins.
    hdr = [c.value for c in header_cells
           if not str(c.value).startswith("__CMP_E1_STATE_V1_")
           and c.value != "__CMP_E2_KEY_V1_TOKEN"]
    state_cells = [c for c in header_cells
                   if str(c.value).startswith("__CMP_E1_STATE_V1_")]
    assert state_cells and all(wb["Comparison"].column_dimensions[c.column_letter].hidden
                               for c in state_cells), state_cells
    wb.close()
    # id columns: PM, #, AENV Row, BENV Row, Status, Diffs, then fields.
    assert hdr[0] == "PM", ("key column must lead the Comparison sheet", hdr)
    assert hdr[6:] == ["County", "Desc"], ("fields = non-key cols in order", hdr)
    os.remove(out)

    # 4) Regression: key_field == 0 reproduces the original field order/lead
    #    column exactly (the byte-identical default path).
    res = run_compare(sc_county, ROWS_A, ROWS_B, False, out, mode="formulas")
    assert res.status == "ok", res.status
    wb = load_workbook(out, data_only=False)
    hdr = [c.value for c in next(wb["Comparison"].iter_rows(max_row=1))
           if not str(c.value).startswith("__CMP_E1_STATE_V1_")
           and c.value != "__CMP_E2_KEY_V1_TOKEN"]
    wb.close()
    assert hdr[0] == "County" and hdr[6:] == ["PM", "Desc"], \
        ("key_field=0 must keep the original layout", hdr)
    os.remove(out)

    print("OK  COMPARE-KEY-IS-FIRST-COLUMN: coarse-key cascade (4 spurious "
          "diff cells) collapses to 0 when keyed on PM; key column leads the "
          "Comparison sheet; key_field=0 layout unchanged.")


if __name__ == "__main__":
    main()
