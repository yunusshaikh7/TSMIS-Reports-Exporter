"""Golden check for the formulas-twin size guard (matrix._comparison_row_count /
matrix._try_formulas) — against REAL workbooks.

v0.18.2 shipped the guard dead: `load_workbook` was only imported inside a
DIFFERENT function, `_comparison_row_count` raised NameError on every call, the
`except Exception` swallowed it at DEBUG and returned None, and None means
"write the twin anyway" — so the skip never fired. No check executed the probe
against a real xlsx, so nothing reddened. This one does, and locks both sides
of the limit:

  * `_comparison_row_count` returns the real data-row count of a real workbook
    (and None for an unreadable path — the fail-open contract);
  * over the row limit `_try_formulas` SKIPS: announces it via events + log,
    never invokes the compare callable, writes no sibling;
  * at/under the limit it RUNS the callable and commits the formulas sibling;
  * CMP-AUD-082: `_settle_formulas_twin` refreshes the twin OR clears a stale
    prior one so a values-only refresh, over-limit skip, or failed refresh can
    never leave an audit-looking `(formulas)` sibling from an older generation.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_formulas_twin_guard.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import matrix
from events import ConsolidateResult, Events
from openpyxl import Workbook

_fail = []


def check(name, cond, detail=""):
    if cond:
        print(f"  ok: {name}")
    else:
        print(f"FAIL: {name}" + (f"\n      {detail}" if detail else ""))
        _fail.append(name)


def _values_workbook(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.append(["Route", "Field"])                 # header
    for i in range(rows):
        ws.append([f"{i:03d}", "x"])
    wb.save(path)


def _fake_compare_call_factory(calls):
    """A compare adapter stand-in: records the call, writes a minimally-valid
    Comparison workbook to the temp path commit_workbook hands it, and returns
    an ok result like the real adapters (commit_workbook passes it through and
    _try_formulas inspects its .status)."""
    def compare_call(tmp_path):
        calls.append(tmp_path)
        wb = Workbook()
        wb.active.title = "Comparison"
        wb.active.append(["h"])
        wb.save(tmp_path)
        return ConsolidateResult(status="ok", message="done",
                                 output_path=str(tmp_path))
    return compare_call


def main():
    tmp = Path(tempfile.mkdtemp(prefix="twin_guard_"))

    # --- the probe reads a REAL workbook (the exact v0.18.2 regression) -----
    values = tmp / "cmp.xlsx"
    _values_workbook(values, rows=5)
    check("_comparison_row_count reads a real workbook (5 data rows)",
          matrix._comparison_row_count(values) == 5,
          f"got {matrix._comparison_row_count(values)!r} — None means the "
          "probe raised and the guard is dead again")
    check("_comparison_row_count is fail-open on an unreadable path",
          matrix._comparison_row_count(tmp / "nope.xlsx") is None)

    logs = []
    events = Events(on_log=logs.append)
    orig_limit = matrix._FORMULAS_TWIN_MAX_ROWS
    try:
        matrix._FORMULAS_TWIN_MAX_ROWS = 3

        # --- over the limit: skip, announce, no sibling, no call -----------
        calls = []
        matrix._try_formulas(_fake_compare_call_factory(calls), values,
                             events=events)
        sibling = matrix._formulas_sibling(values)
        check("over the limit: the compare callable is never invoked",
              calls == [], f"invoked with {calls}")
        check("over the limit: no formulas sibling is written",
              not sibling.exists())
        check("over the limit: the skip is announced via events",
              any("Skipping the live-formulas copy" in m for m in logs),
              f"events got: {logs}")

        # --- at/under the limit: the twin is built and committed ------------
        small = tmp / "small.xlsx"
        _values_workbook(small, rows=3)           # == limit, not over it
        calls2 = []
        matrix._try_formulas(_fake_compare_call_factory(calls2), small,
                             events=events)
        check("under the limit: the compare callable runs",
              len(calls2) == 1, f"calls: {calls2}")
        check("under the limit: the formulas sibling is committed",
              matrix._formulas_sibling(small).exists())

        # S1: Matrix formulas are a derived destination, not the path the user
        # selected.  If that sibling is one of the comparison inputs, the shared
        # commit guard must stop before the comparator runs and preserve it.
        formulas_source = matrix._formulas_sibling(small)
        formulas_prior = formulas_source.read_bytes()
        calls3 = []
        matrix._try_formulas(_fake_compare_call_factory(calls3), small,
                             events=events, source_paths=(formulas_source,))
        check("a formulas sibling that aliases a source is rejected before compare",
              calls3 == [], f"invoked with {calls3}")
        check("the aliased formulas source is preserved byte-for-byte",
              formulas_source.read_bytes() == formulas_prior)

        # --- CMP-AUD-082: a stale (formulas) twin must never outlive its values ---
        # generation. _try_formulas now returns whether a fresh twin was committed,
        # and _settle_formulas_twin refreshes it OR clears a stale prior one so a
        # values-only refresh / over-limit skip / failed refresh can't leave an
        # audit-looking sibling from an older generation beside the newer values copy.
        check("_try_formulas returns True when it commits a fresh twin",
              matrix._try_formulas(_fake_compare_call_factory([]),
                                   tmp / "small.xlsx", events=events) is True)
        check("_try_formulas returns False when it skips over the limit",
              matrix._try_formulas(_fake_compare_call_factory([]), values,
                                   events=events) is False)

        def _seed_twin(vpath):
            matrix._settle_formulas_twin(_fake_compare_call_factory([]), vpath,
                                         True, events=events)
            sib = matrix._formulas_sibling(vpath)
            assert sib.exists(), "seed twin was not written"
            return sib

        # (a) a values-only refresh (do_write False) clears the prior twin.
        va = tmp / "va.xlsx"; _values_workbook(va, rows=3); sib_a = _seed_twin(va)
        matrix._settle_formulas_twin(_fake_compare_call_factory([]), va, False,
                                     events=events)
        check("082: a values-only refresh clears the stale (formulas) twin",
              not sib_a.exists())

        # (b) an over-limit skip clears the prior twin and never writes a new one.
        vb = tmp / "vb.xlsx"; _values_workbook(vb, rows=3); sib_b = _seed_twin(vb)
        _values_workbook(vb, rows=10)                 # now OVER the limit (3)
        callsb = []
        matrix._settle_formulas_twin(_fake_compare_call_factory(callsb), vb, True,
                                     events=events)
        check("082: an over-limit skip clears the stale twin (and wrote nothing)",
              not sib_b.exists() and callsb == [], f"calls={callsb}")

        # (c) a FAILED formulas refresh clears the prior twin (commit kept it).
        vc = tmp / "vc.xlsx"; _values_workbook(vc, rows=3); sib_c = _seed_twin(vc)
        matrix._settle_formulas_twin(
            lambda _p: ConsolidateResult(status="error", message="boom"), vc, True,
            events=events)
        check("082: a failed formulas refresh clears the stale twin",
              not sib_c.exists())

        # (d) clearing is a no-op when no prior twin exists (the common first build).
        vd = tmp / "vd.xlsx"; _values_workbook(vd, rows=3)
        matrix._clear_stale_formulas_twin(vd, events=events)
        check("082: clearing is a no-op when no prior twin exists",
              not matrix._formulas_sibling(vd).exists())

        # (e) a SUCCESSFUL refresh leaves the FRESH twin in place.
        ve = tmp / "ve.xlsx"; _values_workbook(ve, rows=3)
        matrix._settle_formulas_twin(_fake_compare_call_factory([]), ve, True,
                                     events=events)
        check("082: a successful refresh leaves the fresh twin in place",
              matrix._formulas_sibling(ve).exists())

        # RED: the bare _try_formulas path (the pre-fix behavior) LEAVES the stale
        # twin on an over-limit skip — the finding's exact defect.
        vr = tmp / "vr.xlsx"; _values_workbook(vr, rows=3); sib_r = _seed_twin(vr)
        _values_workbook(vr, rows=10)
        wrote = matrix._try_formulas(_fake_compare_call_factory([]), vr, events=events)
        check("082 (red): bare _try_formulas over-limit leaves the stale twin",
              wrote is False and sib_r.exists())
    finally:
        matrix._FORMULAS_TWIN_MAX_ROWS = orig_limit


if __name__ == "__main__":
    print("formulas-twin size guard (real-workbook probe):")
    main()
    if _fail:
        print(f"\n{len(_fail)} check(s) FAILED")
        sys.exit(1)
    print("\nall good")
