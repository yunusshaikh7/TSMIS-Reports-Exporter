"""P5 (R1-N01/S04) -- the shared single-file TSN normalizer factory.

Locks `tsn_library.build_normalized` + the four thin `tsn_load_*` shims that delegate
to it, proving the family DRY is **semantically identical** to the pre-P5 loaders (it
compares the produced workbook's semantics + the result contract, NOT the XLSX ZIP
bytes). For all six cases (Ramp Detail, Intersection Detail, and full + partial of both
summaries) it asserts the FULL signature against an INDEPENDENT frozen oracle (the
sheet/header/alignment literals are hand-written here, NOT re-read from the kwargs the
shim passes to `build_normalized`):

  * every `ConsolidateResult` field (status, message, summary_lines, completion,
    skipped_inputs) -- exact text/order, incl. explicit COMPLETE for detail
    producers and the producer PARTIAL warning for a
    missing summary category;
  * the exact emitted event log line;
  * the sheet title, the header row, and every data row (details: the verbatim rows
    the projection yields; summaries: one [key, count] row per category, distinct fed
    counts, in order, missing -> 0);
  * every header cell's alignment AND font AND fill.

It also locks the shared skeleton's exact strings + branches (deps gate, missing-raw,
overwrite-confirm cancel, parse-error wrap, exact-one admission, `~$` lock skip), the
atomic-save `PermissionError` contract (prior artifact retained, no stray file), and the
P5-A01 ImportError backstop. compare_core is untouched.

Offline: the per-report projection (`tsn_rows_from_raw` / `parse_tsn_pdf`) is
monkeypatched; openpyxl writes real workbooks. Run from the repo root:
    build\\.venv\\Scripts\\python.exe build\\check_tsn_normalizer.py
"""
import contextlib
import os
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

import outcome as oc                       # noqa: E402
import tsn_library                         # noqa: E402
import events as events_mod               # noqa: E402
from openpyxl import load_workbook         # noqa: E402

import tsn_load_ramp_detail as rd_load             # noqa: E402
import tsn_load_intersection_detail as id_load     # noqa: E402
import tsn_load_ramp_summary as rs_load            # noqa: E402
import tsn_load_intersection_summary as is_load    # noqa: E402
import compare_ramp_summary_tsn as rstsn           # noqa: E402
import compare_intersection_summary_tsn as istsn   # noqa: E402

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


@contextlib.contextmanager
def _patch(obj, name, value):
    old = getattr(obj, name)
    setattr(obj, name, value)
    try:
        yield
    finally:
        setattr(obj, name, old)


# --------------------------------------------------------------------------- #
# FROZEN oracle -- hand-written from the v0.18.0 P5 baseline, INDEPENDENT of the
# kwargs the shims pass to build_normalized. A deliberate sheet/header/style change
# must update these literals (the point of a golden tripwire).
# --------------------------------------------------------------------------- #
_FONT = ("Arial", True, 11.0, "00FFFFFF")
_FILL = ("solid", "00305496")
_AL_WRAP = ("center", "center", True)        # ramp / intersection detail
_AL_VCENTER = ("center", "center", None)     # ramp summary
_AL_CENTER = ("center", None, None)          # intersection summary

_RD_SHEET = "Ramp Detail (TSN)"
_ID_SHEET = "Intersection Detail (TSN)"
_RS_SHEET = "Ramp Summary (TSN)"
_IS_SHEET = "Intersection Summary (TSN)"
# Re-blessed to the v0.26.0 v3 layout: the normalized sheet appends the TSN
# District/County sidecar (read by the visual-evidence generator; the comparison
# loader slices it off). Hand-written (the golden tripwire stays independent of
# rd.SHARED_HEADER), so a future header drift is caught.
_RD_SIDECAR = ["TSN District", "TSN County"]
_RD_HEADER = ["Route", "PR", "PM", "Date of Record", "HG", "Area 4", "City Code", "R/U",
              "Description", "Ramp Name", "On/Off", "Ramp Type", "ADT"] + _RD_SIDECAR
# Re-blessed to the v0.22.0 July-2026 layout: the second ML eff-date left the shared
# header, 'Xing Line Lgth' joined at the tail, and the v3 normalized sheet appends the
# TSN District/County sidecar (read by the visual-evidence generator; the comparison
# loader slices it off). Hand-written (the golden tripwire stays independent of
# idt.SHARED_HEADER), so a future header drift is caught.
_ID_SIDECAR = ["TSN District", "TSN County"]
_ID_HEADER = ["Route", "PR", "Route Suffix", "PM", "Date of Record", "HG", "City Code", "R/U",
              "INT Type Eff-Date", "INT Type", "Control Type Eff-Date", "Control Type",
              "Lighting Eff-Date", "Lighting", "ML Eff-Date", "ML Mastarm", "ML Left Chan",
              "ML Right Chan", "ML Traffic Flow", "ML Num Lanes",
              "Description", "Main Line Length", "CS Eff-Date", "CS Mastarm", "CS Left Chan",
              "CS Right Chan", "CS Traffic Flow", "CS Num Lanes", "Int St Eff-Date",
              "Intrte Route", "Intrte PM Prefix", "Intrte Postmile", "Intrte PM Suffix",
              "Xing Line Lgth"] + _ID_SIDECAR
_SUM_HEADER = ["Category", "Count"]
_DEPS_XLSX = "Required components are missing (openpyxl)."
_DEPS_PDF = "Required components are missing (pdfplumber, openpyxl)."
_CANCELLED = "Cancelled. Existing file kept."


def _events():
    logs = []
    return events_mod.Events(on_log=logs.append), logs


def _sig(path):
    """(sheet title, [rows of values], [(align, font, fill) per header cell])."""
    wb = load_workbook(path)
    ws = wb.active
    rows = [[c.value for c in r] for r in ws.iter_rows()]

    def style(c):
        return ((c.alignment.horizontal, c.alignment.vertical, c.alignment.wrap_text),
                (c.font.name, c.font.bold, c.font.size,
                 c.font.color.rgb if c.font.color else None),
                (c.fill.patternType, c.fill.fgColor.rgb))
    head = [style(c) for c in ws[1]]
    title = ws.title
    wb.close()
    return title, rows, head


def test_detail_signatures():
    print("detail loaders -- full frozen signature (result + log + sheet/header/rows + style):")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_tsnnorm_"))
    try:
        # BOTH detail loaders (v3) feed through their OWN tsn_rows_with_dcr (the
        # loader appends the district/county sidecar — Ramp Detail joined in
        # v0.26.0). Each entry: the patch target/factory and how the written rows
        # relate to the fed base rows.
        for loader, label, sheet, header, base_w, target, make_patch, exp_rows_fn in (
                (rd_load, "TSN Ramp Detail", _RD_SHEET, _RD_HEADER,
                 len(_RD_HEADER) - len(_RD_SIDECAR),
                 (rd_load, "tsn_rows_with_dcr"),
                 lambda syn: (lambda _p, s=syn: (list(s), [("01", "DN")] * len(s))),
                 lambda syn: [r + ["01", "DN"] for r in syn]),
                (id_load, "TSN Intersection Detail", _ID_SHEET, _ID_HEADER,
                 len(_ID_HEADER) - len(_ID_SIDECAR),
                 (id_load, "tsn_rows_with_dcr"),
                 lambda syn: (lambda _p, s=syn: (list(s), [("12", "ORA")] * len(s))),
                 lambda syn: [r + ["12", "ORA"] for r in syn])):
            raw = tmp / sheet
            raw.mkdir()
            (raw / "s.xlsx").write_bytes(b"x")
            syn = [["1"] + ["a"] * (base_w - 1),
                   ["1"] + ["b"] * (base_w - 1),
                   ["2"] + ["c"] * (base_w - 1)]          # 3 rows, 2 distinct routes
            out = tmp / f"{sheet}.xlsx"
            ev, logs = _events()
            with _patch(target[0], target[1], make_patch(syn)):
                r = loader.build_into(raw, out, events=ev)
            exp_msg = f"Normalized 3 {label} rows (2 routes)."
            exp_sum = [f"{label}: 3 rows, 2 routes -> {out.name}"]
            check(f"{label}: result fields exact (status/msg/summary/completion/skipped)",
                  (r.status, r.message, list(r.summary_lines), r.completion, r.skipped_inputs)
                  == ("ok", exp_msg, exp_sum, oc.COMPLETE, 0))
            check(f"{label}: event log == one Normalizing line", logs == [f"Normalizing {label}: s.xlsx"])
            title, rows, head = _sig(out)
            check(f"{label}: sheet + header + data rows exact",
                  (title, rows[0], rows[1:]) == (sheet, header, exp_rows_fn(syn)))
            check(f"{label}: every header cell align/font/fill == frozen wrap+blue",
                  all(s == (_AL_WRAP, _FONT, _FILL) for s in head))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _summary_case(loader, cmp_mod, cats, label, sheet, align, total_slug, total_label,
                  miss_idx, out):
    """Drive one summary build with distinct fed counts; return (result, logs, sig,
    expected_rows, total_value)."""
    counts = {slug: i + 1 for i, (_k, slug) in enumerate(cats)}     # distinct per category
    counts[total_slug] = 7777
    if miss_idx is not None:
        counts[cats[miss_idx][1]] = None                            # drop one category
    raw = out.parent / (sheet + "_raw")
    raw.mkdir(parents=True, exist_ok=True)
    (raw / "s.pdf").write_bytes(b"x")
    ev, logs = _events()
    with _patch(cmp_mod, "parse_tsn_pdf", lambda _p, c=counts: dict(c)):
        r = loader.build_into(raw, out, events=ev)
    exp_rows = [[key, (0 if counts.get(slug) is None else counts[slug])]
                for key, slug in cats]
    return r, logs, _sig(out), exp_rows, counts[total_slug]


def test_summary_signatures():
    print("summary loaders -- full frozen signature, complete + producer-partial:")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_tsnnorm_"))
    try:
        for loader, cmp_mod, cats, label, sheet, align, tslug, tlabel in (
                (rs_load, rstsn, list(rstsn._CATEGORIES), "TSN Ramp Summary", _RS_SHEET,
                 _AL_VCENTER, "total_ramps", "Total Number of Ramps"),
                (is_load, istsn, list(istsn._SPEC.categories_for("tsn")),
                 "TSN Intersection Summary", _IS_SHEET, _AL_CENTER, "total_intersections",
                 "Total Intersections")):
            n = len(cats)
            # FULL: no missing category -> complete / 0 skipped
            out = tmp / f"{sheet}_full.xlsx"
            r, logs, (title, rows, head), exp_rows, total = _summary_case(
                loader, cmp_mod, cats, label, sheet, align, tslug, tlabel, None, out)
            exp_sum = [f"{label}: {n} categories -> {out.name}", f"{tlabel}: {total}"]
            check(f"{label} FULL: result fields exact",
                  (r.status, r.message, list(r.summary_lines), r.completion, r.skipped_inputs)
                  == ("ok", f"Normalized {label} ({n} categories).", exp_sum, oc.COMPLETE, 0))
            check(f"{label} FULL: event log", logs == [f"Normalizing {label}: s.pdf"])
            check(f"{label} FULL: sheet + header + every [key,count] row exact",
                  (title, rows[0], rows[1:]) == (sheet, _SUM_HEADER, exp_rows))
            check(f"{label} FULL: header cell align/font/fill == frozen",
                  all(s == (align, _FONT, _FILL) for s in head))
            # PARTIAL: drop category 0 -> partial + skipped_inputs=1 + the warning line first
            out = tmp / f"{sheet}_part.xlsx"
            r, logs, (title, rows, head), exp_rows, total = _summary_case(
                loader, cmp_mod, cats, label, sheet, align, tslug, tlabel, 0, out)
            warn = (f"⚠ INCOMPLETE — 1 category not found in the PDF: {cats[0][0]}")
            exp_sum = [warn, f"{label}: {n} categories -> {out.name}", f"{tlabel}: {total}"]
            check(f"{label} PARTIAL: result fields exact (partial + 1 skipped + warning)",
                  (r.status, r.message, list(r.summary_lines), r.completion, r.skipped_inputs)
                  == ("ok", f"Normalized {label} ({n} categories).", exp_sum, oc.PARTIAL, 1))
            check(f"{label} PARTIAL: dropped category row is [key, 0]",
                  rows[1] == [cats[0][0], 0])
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_shared_skeleton():
    print("the shared factory skeleton -- exact strings + branches:")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_tsnnorm_"))
    try:
        rows = [["1"] + ["a"] * (len(_RD_HEADER) - len(_RD_SIDECAR) - 1)]
        raw = tmp / "rd"
        raw.mkdir()
        # deps gate -> exact friendly deps message
        with _patch(rd_load, "_DEPS_OK", False):
            r = rd_load.build_into(raw, tmp / "x.xlsx", events=_events()[0])
        check("deps_ok False -> error with the EXACT deps message",
              r.status == "error" and r.message == _DEPS_XLSX)
        # missing raw -> exact 'No raw ... found' message naming the report + raw_dir
        r = rd_load.build_into(raw, tmp / "x.xlsx", events=_events()[0])
        check("missing raw -> exact 'No raw TSN Ramp Detail .xlsx found' + raw_dir + hint",
              r.status == "error"
              and r.message == (f"No raw TSN Ramp Detail .xlsx found in:\n{raw}\n\n"
                                "Import the statewide 'TSAR - RAMPS DETAIL' TSN export first."))
        # overwrite-confirm False -> exact cancelled string, prior file untouched
        (raw / "s.xlsx").write_bytes(b"x")
        out = tmp / "out.xlsx"
        out.write_bytes(b"existing")
        with _patch(rd_load, "tsn_rows_with_dcr",
                    lambda _p: (list(rows), [("01", "DN")] * len(rows))):
            r = rd_load.build_into(raw, out, events=_events()[0], confirm_overwrite=lambda _p: False)
        check("existing out + confirm False -> exact cancelled string, prior bytes kept",
              r.status == "cancelled" and r.message == _CANCELLED and out.read_bytes() == b"existing")
        # parse error -> exact 'Could not read ...' wrap
        with _patch(rd_load, "tsn_rows_with_dcr",
                    lambda _p: (_ for _ in ()).throw(ValueError("bad parse"))):
            r = rd_load.build_into(raw, tmp / "y.xlsx", events=_events()[0])
        check("projection raises -> exact 'Could not read s.xlsx: ValueError: bad parse'",
              r.status == "error" and r.message == "Could not read s.xlsx: ValueError: bad parse")
        # zero-row projection -> error, nothing written, prior normalized kept
        # (a TSN layout change must not become an "ok" EMPTY library that turns
        # every comparison row into "Only in TSMIS").
        prev = tmp / "prev.xlsx"
        prev.write_bytes(b"prior-normalized")
        with _patch(rd_load, "tsn_rows_with_dcr", lambda _p: ([], [])):
            r = rd_load.build_into(raw, prev, events=_events()[0],
                                   confirm_overwrite=lambda _p: True)
        check("zero-row projection -> error suggesting a layout change",
              r.status == "error" and "produced 0 rows" in r.message)
        check("...previous normalized bytes kept",
              prev.read_bytes() == b"prior-normalized")
        # atomic-save PermissionError -> exact friendly message; prior artifact retained; no .tmp-* left.
        # Drive the REAL artifact_store.atomic_save (it saves the workbook to a `.tmp-*` sibling of
        # out_path, then os.replace); force os.replace to raise -> atomic_save removes its temp and
        # re-raises -> the factory's `except PermissionError` returns the friendly result. Exercising
        # the real save path (not a stubbed atomic_save) proves the temp cleanup AND avoids the
        # write-only-workbook atexit noise (the workbook IS saved before os.replace fails). (P5-R01 r2.)
        import artifact_store
        od = tmp / "savedir"
        od.mkdir()
        out = od / "keep.xlsx"
        out.write_bytes(b"prior")

        def _boom_replace(*a, **k):
            raise PermissionError(13, "destination open in Excel")
        with _patch(rd_load, "tsn_rows_with_dcr",
                    lambda _p: (list(rows), [("01", "DN")] * len(rows))), \
             _patch(artifact_store.os, "replace", _boom_replace):
            r = rd_load.build_into(raw, out, events=_events()[0])
        check("atomic-save PermissionError (real os.replace) -> exact 'probably open in Excel' message",
              r.status == "error" and r.message == (f"Could not save {out.name}.\n\n"
              "The file is probably open in Excel. Close it and try again."))
        check("...prior output bytes retained + atomic_save left no .tmp-* sibling in the output dir",
              out.read_bytes() == b"prior" and sorted(p.name for p in od.iterdir()) == ["keep.xlsx"])
        # P5-A01: a partial openpyxl (workbook symbol missing) -> friendly deps result, not a crash
        with _patch(rd_load, "tsn_rows_with_dcr",
                    lambda _p: (list(rows), [("01", "DN")] * len(rows))), \
             _patch(tsn_library, "_write_normalized_workbook",
                    lambda *a, **k: (_ for _ in ()).throw(ImportError("no WriteOnlyCell"))):
            r = rd_load.build_into(raw, tmp / "z.xlsx", events=_events()[0])
        check("workbook-symbol ImportError -> backstopped to the deps message (P5-A01)",
              r.status == "error" and r.message == _DEPS_XLSX)
        # Two ordinary candidates are ambiguous regardless of mtime; Excel ~$ locks
        # do not count as candidates.
        raw2 = tmp / "rd2"
        raw2.mkdir()
        for nm, mt in (("a.xlsx", 3000.0), ("z.xlsx", 1000.0), ("~$lock.xlsx", 9000.0)):
            f = raw2 / nm
            f.write_bytes(b"x")
            os.utime(f, (mt, mt))
        seen = {}

        def _capture(p):
            seen["path"] = p
            return list(rows), [("01", "DN")] * len(rows)
        with _patch(rd_load, "tsn_rows_with_dcr", _capture):
            ambiguous = rd_load.build_into(raw2, tmp / "z2.xlsx", events=_events()[0])
        check("two ordinary raws are rejected (no newest-by-mtime selection)",
              ambiguous.status == "error" and "Found 2 ordinary matching files" in ambiguous.message
              and not seen)
        (raw2 / "a.xlsx").unlink()
        with _patch(rd_load, "tsn_rows_with_dcr", _capture):
            accepted = rd_load.build_into(raw2, tmp / "z2.xlsx", events=_events()[0])
        check("one ordinary raw accepted while ~$ owner lock is ignored",
              accepted.status == "ok" and Path(seen.get("path", "")).name == "z.xlsx")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main():
    test_detail_signatures()
    test_summary_signatures()
    test_shared_skeleton()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL TSN-NORMALIZER CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
