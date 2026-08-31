"""HF-06 / PCOA-FINAL-011: the Highway Sequence PDF-vs-Excel self check must
normalize the equate RELATION and must still report every real divergence.

A postmile equation is ONE fact the two renders spell differently by design.
The print writes the TSN convention - an annotation line "EQUATES TO <label>"
at the realignment postmile with HG / FT / Distance / suffix structurally
blank, then the equated postmile's own line carrying the "E" suffix. The Excel
export has no annotation convention: it folds the marker, the label, the
segment's flags and (about a quarter of the time) the "E" itself onto the
realignment record.

On the frozen 2026-07-23 statewide pull that spelling published 3,707 of the
self check's 3,714 differing cells. The owner ruled 2026-07-26 that they are
not discrepancies, so the relation is canonicalized before comparing.

The rule has to be PAIR-AWARE: the "E" genuinely sits on a different ROW on
each side, so no cell-by-cell rule could close the PM Suffix column. It must
also be narrow, because the very same class contains real divergences - an
"E" that only one render carries anywhere in the relation is a genuine
disagreement about whether the marker exists at all, and seven of those
survive statewide.

Every case below drives the SHIPPED adapter (`TSMIS_PDF_VS_EXCEL.compare`)
over real per-route-shaped workbooks with the PDF-source marker, and reads the
verdict back out of the written VALUES workbook - not the internal helper.

CI-safe: pure Python fixtures, no local data.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_highway_sequence_equate.py
"""
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import Workbook, load_workbook          # noqa: E402

import compare_highway_sequence_pdf as hslp           # noqa: E402
import compare_highway_sequence_tsn as hsl            # noqa: E402
from compare_core import _DIFF_MARK                   # noqa: E402
from events import Events                             # noqa: E402
from pdf_table_lib import write_pdf_source_marker     # noqa: E402

failures = []
HEADER = ["Route"] + [("" if h is None else h) for h in hsl._TSMIS_HEADER[1:]]
# Consolidated Highway Sequence positions (Route prepended).
ROUTE, COUNTY, CITY, PREFIX, PM, SUFFIX, HG, FT, DIST, DESC = range(10)


def check(label, condition, detail=""):
    print(("OK   " if condition else "FAIL ") + label
          + (f" - {detail}" if detail and not condition else ""))
    if not condition:
        failures.append(label)


def row(county, prefix, pm, suffix="", hg="", ft="", dist="", desc="",
        route="001", city=""):
    out = [""] * len(HEADER)
    out[ROUTE], out[COUNTY], out[CITY] = route, county, city
    out[PREFIX], out[PM], out[SUFFIX] = prefix, pm, suffix
    out[HG], out[FT], out[DIST], out[DESC] = hg, ft, dist, desc
    return out


def write_wb(path, rows, marked=False):
    wb = Workbook()
    ws = wb.active
    ws.title = hsl.TSMIS_SHEET
    ws.append(list(HEADER))
    for r in rows:
        ws.append(list(r))
    if marked:
        write_pdf_source_marker(wb)
    wb.save(path)
    wb.close()


def compare(tmp, print_rows, excel_rows, tag):
    """Run the shipped self check and return (result, differing display cells)."""
    a, b = tmp / f"{tag}_pdf.xlsx", tmp / f"{tag}_xls.xlsx"
    out = tmp / f"{tag}_cmp.xlsx"
    write_wb(a, print_rows, marked=True)
    write_wb(b, excel_rows)
    result = hslp.TSMIS_PDF_VS_EXCEL.compare(
        str(a), str(b), str(out), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    if result.status != "ok":
        return result, None
    wb = load_workbook(out, read_only=True, data_only=True)
    try:
        cells = [value
                 for line in wb["Comparison"].iter_rows(values_only=True)
                 for value in line
                 if isinstance(value, str) and _DIFF_MARK in value]
    finally:
        wb.close()
    return result, cells


def counts(result):
    c = result.comparison_outcome.counts
    return c.differing_cells, c.differing_rows, c.side_a_only_rows, c.side_b_only_rows


# --------------------------------------------------------------------------- #
# The measured route-001 shape: ORA R018.540 "EQUATES TO END R REALIGNMENT"
# equating to ORA 018.530. Print: annotation blank-flagged, "E" on the target.
# Excel: label alone, the segment's D/H repeated, "E" on the annotation.
# --------------------------------------------------------------------------- #
def print_pair(label="EQUATES TO END R REALIGNMENT"):
    return [row("ORA", "R", "018.446", hg="U", ft="H", dist="000.094",
                desc="DOVER DR-RT/BAYSHORE-LT"),
            row("ORA", "R", "018.540", desc=label),
            row("ORA", "", "018.530", suffix="E", hg="D", ft="H",
                dist="001.267"),
            row("ORA", "", "018.833", hg="D", ft="I",
                desc="BALBOA BAY CLUB - LT")]


def excel_pair(label="END R REALIGNMENT", target_hg="D", target_ft="H",
               annotation_suffix="E", target_suffix=""):
    return [row("ORA", "R", "018.446", hg="U", ft="H", dist="000.094",
                desc="DOVER DR-RT/BAYSHORE-LT"),
            row("ORA", "R", "018.540", suffix=annotation_suffix, hg="D",
                ft="H", desc=label),
            row("ORA", "", "018.530", suffix=target_suffix, hg=target_hg,
                ft=target_ft, dist="001.267"),
            row("ORA", "", "018.833", hg="D", ft="I",
                desc="BALBOA BAY CLUB - LT")]


tmp = Path(tempfile.mkdtemp(prefix="tsmis_hsl_equate_"))
try:
    # (a) the by-design spelling alone -> ZERO differences ------------------
    result, cells = compare(tmp, print_pair(), excel_pair(), "a")
    check("(a) the measured equate spelling reports ZERO differing cells",
          result.status == "ok" and cells == [],
          f"{result.status}: {cells}")
    check("(a) ... and zero differing rows, with no fabricated one-sided row",
          result.status == "ok" and counts(result) == (0, 0, 0, 0),
          str(counts(result)) if result.status == "ok" else result.message)
    check("(a) ... and the run says EVERYTHING MATCHES",
          any("EVERYTHING MATCHES" in line for line in result.summary_lines),
          str(result.summary_lines[:1]))

    # The four columns the class occupied must ALL close, PM Suffix included:
    # that is what proves the rule is pair-aware rather than per-cell.
    check("(a) ... closing PM Suffix too - the 'E' moved ROW, so a per-cell "
          "rule could not have",
          all(_DIFF_MARK not in (value or "") for value in (cells or [])))

    # (b) a real Description LABEL change still reports ---------------------
    result, cells = compare(tmp, print_pair(),
                            excel_pair(label="END X REALIGNMENT"), "b")
    check("(b) a real Description label change is STILL a difference",
          result.status == "ok" and len(cells) == 1
          and "END R REALIGNMENT" in cells[0] and "END X REALIGNMENT" in cells[0],
          f"{result.status}: {cells}")

    # (c) a real HG / FT change on the PARTNER row still reports ------------
    result, cells = compare(tmp, print_pair(), excel_pair(target_hg="U"), "c")
    check("(c) a real HG change on the partner (target) row is STILL a "
          "difference",
          result.status == "ok" and len(cells) == 1 and "D" in cells[0]
          and "U" in cells[0], f"{result.status}: {cells}")
    result, cells = compare(tmp, print_pair(), excel_pair(target_ft="R"), "c2")
    check("(c) a real FT change on the partner (target) row is STILL a "
          "difference",
          result.status == "ok" and len(cells) == 1, f"{result.status}: {cells}")

    # (d) an "E" only ONE render carries, anywhere in the pair --------------
    result, cells = compare(tmp, print_pair(),
                            excel_pair(annotation_suffix=""), "d")
    check("(d) an 'E' the print carries and the export does not, anywhere in "
          "the relation, is STILL a difference",
          result.status == "ok" and len(cells) == 1 and "E" in cells[0],
          f"{result.status}: {cells}")
    # ... and the reverse direction (the statewide route-580 case).
    print_no_marker = print_pair()
    print_no_marker[2][SUFFIX] = ""
    result, cells = compare(tmp, print_no_marker, excel_pair(), "d2")
    check("(d) an 'E' the export carries and the print does not is STILL a "
          "difference",
          result.status == "ok" and len(cells) == 1 and "E" in cells[0],
          f"{result.status}: {cells}")

    # (e) a county/route-boundary relation and a DELAYED target ------------
    #     The target is in the next COUNTY (its own key), several rows on.
    boundary_print = [
        row("SD", "", "000.553", desc="EQUATES TO"),
        row("SD", "R", "000.000", hg="R", ft="H", dist="000.000",
            desc="JCT 5/15 SEP 57-438"),
        row("SD", "R", "000.122", hg="R", ft="H", dist="000.019",
            desc="WABASH VIADUCT 57-732"),
        row("SD", "R", "000.395", hg="R", ft="H", dist="000.010",
            desc="END RT INDEP ALIGN"),
        row("RIV", "R", "000.000", suffix="E", hg="L", ft="H", dist="000.000",
            desc="BEGIN LT INDEP ALIGN"),
    ]
    boundary_excel = [
        row("SD", "", "000.553", suffix="E", hg="L", ft="H",
            desc="PM EQUATION"),
        row("SD", "R", "000.000", hg="R", ft="H", dist="000.000",
            desc="JCT 5/15 SEP 57-438"),
        row("SD", "R", "000.122", hg="R", ft="H", dist="000.019",
            desc="WABASH VIADUCT 57-732"),
        row("SD", "R", "000.395", hg="R", ft="H", dist="000.010",
            desc="END RT INDEP ALIGN"),
        row("RIV", "R", "000.000", hg="L", ft="H", dist="000.000",
            desc="BEGIN LT INDEP ALIGN"),
    ]
    result, cells = compare(tmp, boundary_print, boundary_excel, "e")
    check("(e) a DELAYED target across a county boundary normalizes, and the "
          "bare 'EQUATES TO' pairs with the export's 'PM EQUATION'",
          result.status == "ok" and cells == [], f"{result.status}: {cells}")

    # ... and the same relation with a genuinely changed label still reports.
    boundary_excel_changed = [list(r) for r in boundary_excel]
    boundary_excel_changed[0][DESC] = "ROUTE BREAK"
    result, cells = compare(tmp, boundary_print, boundary_excel_changed, "e2")
    check("(e) ... but an export label that is NOT the label-less marker is "
          "still compared",
          result.status == "ok" and len(cells) == 1 and "ROUTE BREAK" in cells[0],
          f"{result.status}: {cells}")

    # (f) the rule is scoped: it fires only where the PRINT declared an equate
    ordinary = [row("ORA", "R", "018.540", hg="D", ft="H",
                    desc="END R REALIGNMENT")]
    ordinary_changed = [row("ORA", "R", "018.540", hg="U", ft="H",
                            desc="END R REALIGNMENT")]
    result, cells = compare(tmp, ordinary, ordinary_changed, "f")
    check("(f) a row that is NOT an equate annotation keeps every cell "
          "compared",
          result.status == "ok" and len(cells) == 1, f"{result.status}: {cells}")

    # ... and a print row whose flags are NOT blank is not this class at all,
    # so its "EQUATES TO" text is compared verbatim (the rule fails OPEN).
    flagged_print = [row("ORA", "R", "018.540", hg="D", ft="H",
                         desc="EQUATES TO END R REALIGNMENT")]
    flagged_excel = [row("ORA", "R", "018.540", hg="D", ft="H",
                         desc="END R REALIGNMENT")]
    result, cells = compare(tmp, flagged_print, flagged_excel, "f2")
    check("(f) an 'EQUATES TO' print row carrying flags is NOT the annotation "
          "class - it stays compared verbatim (fails OPEN)",
          result.status == "ok" and len(cells) == 1
          and "EQUATES TO" in cells[0], f"{result.status}: {cells}")

    # ... and the pure canonicalizer is a no-op where the print declares none.
    plain = [row("ORA", "R", "018.540", hg="D", ft="H", desc="A LANDMARK")]
    left, right, relations = hslp.canonicalize_equate_pair(
        [list(r) for r in plain], [list(r) for r in plain])
    check("(f) canonicalize_equate_pair is a NO-OP with no print annotation",
          relations == 0 and left == plain and right == plain,
          f"relations={relations}")

    # ... and the vs-TSN flavor never canonicalizes: it is a different loader,
    # and its equate disclosure is the pre-existing one.
    check("(f) the vs-TSN flavor does not normalize equates",
          hslp.TSMIS_PDF_VS_TSN._same_source is False
          and hslp.TSMIS_PDF_VS_TSN._schema_for("x").disclosure_notes == ())

    # The disclosure a reader needs to understand why the number moved.
    schema = hslp.TSMIS_PDF_VS_EXCEL._schema_for(None, {"relations": 1119})
    note = schema.disclosure_notes[0]()
    check("the Summary discloses the normalized class and its relation count",
          "1,119" in note and "EQUATES TO" in note and "normalized" in note,
          note[:120])
    notes_wb = Workbook()
    schema.legend_writer(notes_wb)
    notes_text = [cell.value for cell in notes_wb["Notes"]["A"]]
    notes_wb.close()
    check("... and the Notes sheet carries the same resolved line",
          note in notes_text,
          f"last note: {str(notes_text[-1])[:120]!r}")
    check("... and the Notes sheet says what the rule does NOT hide",
          any(isinstance(line, str) and "does NOT hide" in line
              for line in notes_text))
finally:
    shutil.rmtree(tmp, ignore_errors=True)

if failures:
    print(f"\nFAILED {len(failures)} check(s):")
    for item in failures:
        print("  -", item)
    sys.exit(1)
print("\nOK  COMPARE-HIGHWAY-SEQUENCE-EQUATE: the PDF-vs-Excel self check "
      "normalizes the by-design equate relation to zero differing cells - "
      "pair-aware, so PM Suffix closes too - while a changed label, a real "
      "HG/FT change on the partner row, and a one-sided 'E' anywhere in the "
      "relation all still report; the rule fires only where the print "
      "declared an equate, and the relation count is disclosed.")
