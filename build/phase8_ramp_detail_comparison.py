#!/usr/bin/env python3
"""Independent Stage-8 Ramp Detail four-source comparison oracle.

This module deliberately imports no production parser, normalizer, comparator,
schema, consolidator, evidence adapter, or application constant.  It reads the
authoritative TSMIS XLSX/PDF pair and TSN XLSX source through separately declared
contracts, then binds the already accepted TSN XLSX/PDF and Stage-6 evidence.

The owner-approved comparison identity is exactly
``(Route, County, norm_pm(PM))``.  District, PR, dates, HG, Area 4, City, R/U,
and Description remain asserted fields.  TSMIS-PDF-vs-TSN additionally asserts
On/Off and Ramp Type.  PM_SFX is conserved as a separate TSN source claim; it is
not silently inferred into the key.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from contextlib import contextmanager
from dataclasses import asdict, dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import gc
import hashlib
import itertools
import json
import logging
import math
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import tempfile
import time
from typing import Iterable, Sequence
import zipfile

import openpyxl
from openpyxl import load_workbook
import pdfplumber
from pypdf import PdfReader

logging.getLogger("pdfminer").setLevel(logging.ERROR)


REPO_ROOT = Path(__file__).resolve().parent.parent
GENERATOR_PATH = Path(__file__).resolve()
PRODUCT_HELPER_PATH = GENERATOR_PATH.with_name(
    "phase8_ramp_detail_product_witness.py")
SELF_GATE_PATH = GENERATOR_PATH.with_name(
    "check_phase8_ramp_detail_comparison.py")
DEFAULT_WORK_ROOT = REPO_ROOT / "tmp" / "phase8-ramp-detail-work"
SOURCE_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9"
    r"\2026-07-09 ssor-prod")
DEFAULT_TSMIS_XLSX_ROOT = SOURCE_ROOT / "ramp_detail"
DEFAULT_TSMIS_PDF_ROOT = SOURCE_ROOT / "ramp_detail_pdf"
DEFAULT_TSN_XLSX = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\ramp_detail\raw"
    r"\TSAR - RAMPS DETAIL_TSN_11.04.2025IT.xlsx")
DEFAULT_TSN_PDF = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\ramp_detail\pdf"
    r"\Ramp Detail Statewide_TSN.pdf")
R7_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd\phase4_tsn_rebaseline")
DEFAULT_TSN_NORMALIZED = (
    R7_ROOT / "raw-2026-07-12-r7" / "ramp_detail" / "consolidated" /
    "tsn_ramp_detail_normalized.xlsx")
DEFAULT_TSN_NORMALIZED_SIDECAR = Path(str(DEFAULT_TSN_NORMALIZED) + ".outcome.json")
DEFAULT_STAGE6_RESULT = R7_ROOT / "phase6_ramp_detail_conservation_r7_reissued.json"
DEFAULT_STAGE6_ACCEPTANCE = Path(str(DEFAULT_STAGE6_RESULT) + ".acceptance.json")
DEFAULT_TSN_CROSS_FORMAT = (
    R7_ROOT / "raw-2026-07-12-r1" / "ramp_cross_format_oracle-v3.json")

TREE_BINDINGS = {
    "tsmis_xlsx": {
        "files": 126,
        "bytes": 7_858_480,
        "manifest_sha256": "7c10fbf6b996a8a9fbb0e8c8c30d8d2dac0a80c0befb7c12bdeb0151f7ff7489",
        "suffix": ".xlsx",
    },
    "tsmis_pdf": {
        "files": 126,
        "bytes": 12_792_211,
        "manifest_sha256": "6e8a2b669148738344a0173cca52a16884b972cba4679ba6446547ce8286c4c9",
        "suffix": ".pdf",
    },
}
FILE_BINDINGS = {
    "tsn_xlsx": {
        "bytes": 1_590_431,
        "sha256": "3e0c552a0a130db07275eed776a05f2a3bd0b438b53eb33ceec54bdd9c722856",
    },
    "tsn_pdf": {
        "bytes": 1_384_895,
        "sha256": "0d1e31054e8f866de3be924ba350a5bd77f9230d453e58d761dea079f4505a49",
    },
    "tsn_normalized": {
        "bytes": 1_009_829,
        "sha256": "c121a9ca1bed2fad00bfc4b08bfc68fa01cd46da436d6bffa699c5579bb4f5f1",
    },
    "tsn_normalized_sidecar": {
        "bytes": 910,
        "sha256": "980ccd48f0c15438547b32fbb31050329fd11c94a1f199156c3b3a664f82f5b0",
    },
    "stage6_result": {
        "bytes": 64_727,
        "sha256": "3386ca24768c7182ad79069c80d2d4e103a192bb6af6a6c8b1bcba7c6c1ea1bd",
    },
    "stage6_acceptance": {
        "bytes": 5_941,
        "sha256": "2c346786f27eab3999f225e5821ddf7b08296faf006f5ab2738293a40ccca6cb",
    },
    "tsn_cross_format": {
        "bytes": 72_950,
        "sha256": "47383b5d00ed4b72fa72ed711d165c0ec633d2d7c8f86edd695f4f0a2e886ed1",
    },
}

TSMIS_SHEET = "TSAR - Ramp Detail"
TSMIS_HEADER = (
    "Location", None, "PM", "Date of Record", None, "HG", "Area 4",
    None, "City Code", "R/U", "Description")
TSMIS_PDF_HEADER = (*TSMIS_HEADER, "On/Off", "Ramp Type")
TSN_SHEET = "Sheet 1"
TSN_HEADER = (
    "RAM_CONNECTION_ID", "RAMP_NANE", "LOCATION", "PR", "PM", "PM_SFX",
    "DATE_OF_RECORD", "HG", "AREA_4", "CITY_CODE", "POP", "ON_OFF",
    "ADT_EFF_YEAR", "ADT", "RAMP_TYPE", "EFF_DATE", "DESCRIPTION",
    "SEG_ORDER_ID")
TSN_NORMALIZED_SHEET = "Ramp Detail (TSN)"
TSN_NORMALIZED_HEADER = (
    "Route", "PR", "PM", "Date of Record", "HG", "Area 4", "City Code",
    "R/U", "Description", "Ramp Name", "On/Off", "Ramp Type", "ADT",
    "TSN District", "TSN County")

BASE_ASSERTED_FIELDS = (
    "district", "pr", "record_date", "hg", "area4", "city", "ru",
    "description")
PDF_TSN_ASSERTED_FIELDS = (*BASE_ASSERTED_FIELDS, "onoff", "ramp_type")
EXCEL_TSN_CONTEXT_FIELDS = (
    "ramp_name", "onoff", "ramp_type", "adt", "pm_suffix",
    "adt_eff_year", "effective_date")
PDF_TSN_CONTEXT_FIELDS = (
    "ramp_name", "adt", "pm_suffix", "adt_eff_year", "effective_date")

LOCATION_RE = re.compile(
    r"^(?P<district>\d{2})-(?P<county>[A-Z]{2,3})-(?P<route>\d{3}[A-Z]?)$")
ROUTE_NAME_RE = re.compile(r"route[_ -]*(\d+[A-Za-z]?)", re.IGNORECASE)
ROUTE_RE = re.compile(r"^(\d+)([A-Z]?)$")
TSMIS_DESCRIPTION_RE = re.compile(
    r"^\s*(?P<route>\d+[A-Za-z]?)\s*/(?P<description>.*)$", re.DOTALL)
WS_RE = re.compile(r"\s+")
PDF_PM_RE = re.compile(r"^\d{3}\.\d{3}$")
PDF_PREFIXES = frozenset("CDGHLMNRST")
PDF_HEADER_TOKENS = (
    "LOCATION", "PM", "RECORD", "AREA", "CITY", "CODE", "DESCRIPTION")
NULL_DESCRIPTION = "NO RAMP LINEAR EVENT"

EXPECTED_SOURCE_DIGESTS = {
    "tsmis_excel": {
        "rows": 15_216,
        "route_count": 126,
        "unique_keys": 15_215,
        "duplicate_groups": 1,
        "ordered_row_sha256": (
            "d5d00f705b154de580bd6e02945739efe0b1841214b0d4ee3993cebf7e798c1a"),
        "raw_representation_sha256": (
            "5ebe7c2c8a7fc866edd8d769b87988b6b4c43ac3c18db2729ff204754498abbd"),
    },
    "tsmis_pdf": {
        "rows": 15_216,
        "route_count": 126,
        "unique_keys": 15_215,
        "duplicate_groups": 1,
        "ordered_row_sha256": (
            "93cce3b1b06edfbf480d938b50ebfc5db634e9e807ca3f307fd6014877b12887"),
        "raw_representation_sha256": (
            "9336b7ddb57ad09813cbee05cc3ecca806b1f7e9bf23805d3d36449eda80335c"),
    },
    "tsn_excel": {
        "rows": 15_410,
        "route_count": 126,
        "unique_keys": 15_410,
        "duplicate_groups": 0,
        "ordered_row_sha256": (
            "a7ca66f787c9631293ca68ae099c686a8942240050a83b7cf57809c00cad2f03"),
        "raw_representation_sha256": (
            "994aa5ef2afd0d9e3504cc4b6ecc036e6e0d0fa69606a559950ed3b41b9f379b"),
    },
    "tsn_normalized": {
        "rows": 15_410,
        "route_count": 126,
        "unique_keys": 15_410,
        "duplicate_groups": 0,
        "ordered_row_sha256": (
            "0ec4df15602abc4326daa062cafdd49ad093adafd2a5dcfa90a1b9f3befc98b0"),
        "raw_representation_sha256": (
            "ce11c9efc00ff40c3fee2d1c4f45901b23b2304aaf817ade51cc605ac75e0d93"),
    },
}

EXPECTED_COMPARISONS = {
    "tsmis_excel_vs_tsn": {
        "left_rows": 15_216, "right_rows": 15_410,
        "paired_rows": 15_212, "left_only_rows": 4,
        "right_only_rows": 198, "differing_rows": 741,
        "identical_rows": 14_471, "differing_cells": 847,
        "asserted_cells": 121_696,
        "per_field": {
            "district": 1, "pr": 0, "record_date": 15, "hg": 364,
            "area4": 58, "city": 156, "ru": 68, "description": 185,
        },
        "paired_ledger_sha256": (
            "7bd713435c4d20d7ea0ffccfc23c26d1f6ad23418b5cde60e24489594ff33e73"),
    },
    "tsmis_pdf_vs_tsn": {
        "left_rows": 15_216, "right_rows": 15_410,
        "paired_rows": 15_212, "left_only_rows": 4,
        "right_only_rows": 198, "differing_rows": 774,
        "identical_rows": 14_438, "differing_cells": 998,
        "asserted_cells": 152_120,
        "per_field": {
            "district": 1, "pr": 0, "record_date": 15, "hg": 364,
            "area4": 58, "city": 156, "ru": 68, "description": 181,
            "onoff": 95, "ramp_type": 60,
        },
        "paired_ledger_sha256": (
            "c7edcc516cc10ef7e687c106d1d8b0de28a4811629e2cf34a921f0583b6b2310"),
    },
    "tsmis_pdf_vs_excel": {
        "left_rows": 15_216, "right_rows": 15_216,
        "paired_rows": 15_216, "left_only_rows": 0,
        "right_only_rows": 0, "differing_rows": 4,
        "identical_rows": 15_212, "differing_cells": 4,
        "asserted_cells": 121_728,
        "per_field": {
            "district": 0, "pr": 0, "record_date": 0, "hg": 0,
            "area4": 0, "city": 0, "ru": 0, "description": 4,
        },
        "paired_ledger_sha256": (
            "050fb352d69565f54bf3df07aa80c251ea2ba584b4d4034085e3d587fd01938d"),
    },
    "raw_vs_normalized": {
        "left_rows": 15_410, "right_rows": 15_410,
        "paired_rows": 15_410, "left_only_rows": 0,
        "right_only_rows": 0, "differing_rows": 15,
        "identical_rows": 15_395, "differing_cells": 15,
        "asserted_cells": 184_920,
        "per_field": {
            "district": 0, "pr": 0, "record_date": 0, "hg": 0,
            "area4": 0, "city": 0, "ru": 0, "description": 15,
            "ramp_name": 0, "onoff": 0, "ramp_type": 0, "adt": 0,
        },
        "paired_ledger_sha256": (
            "4a9de9a5369e4f104f40ead979becec6ef1d39f0b9d92a738d40b998e4ada131"),
    },
}

EXPECTED_PRODUCT_COMPARISONS = {
    "excel_vs_tsn_raw": {
        "side_a": "TSMIS", "side_b": "TSN",
        "both_rows": 15_212, "side_a_only_rows": 4,
        "side_b_only_rows": 198, "differing_rows": 750,
        "differing_cells": 861,
        "per_field": {
            "PR": 0, "Date of Record": 15, "HG": 364, "Area 4": 58,
            "City Code": 156, "R/U": 68, "Description": 200,
            "Ramp Name": 0, "On/Off": 0, "Ramp Type": 0, "ADT": 0,
        },
        "ordered_rows_sha256": (
            "8a28a42cd63689a6672f58367be8a0dca2764075477bdb22a62a03778fddc913"),
    },
    "excel_vs_tsn_normalized": {
        "side_a": "TSMIS", "side_b": "TSN",
        "both_rows": 15_212, "side_a_only_rows": 4,
        "side_b_only_rows": 198, "differing_rows": 750,
        "differing_cells": 861,
        "per_field": {
            "PR": 0, "Date of Record": 15, "HG": 364, "Area 4": 58,
            "City Code": 156, "R/U": 68, "Description": 200,
            "Ramp Name": 0, "On/Off": 0, "Ramp Type": 0, "ADT": 0,
        },
        "ordered_rows_sha256": (
            "8a28a42cd63689a6672f58367be8a0dca2764075477bdb22a62a03778fddc913"),
    },
    "pdf_vs_tsn_raw": {
        "side_a": "TSMIS (PDF)", "side_b": "TSN",
        "both_rows": 15_212, "side_a_only_rows": 4,
        "side_b_only_rows": 198, "differing_rows": 783,
        "differing_cells": 1_012,
        "per_field": {
            "PR": 0, "Date of Record": 15, "HG": 364, "Area 4": 58,
            "City Code": 156, "R/U": 68, "Description": 196,
            "Ramp Name": 0, "On/Off": 95, "Ramp Type": 60, "ADT": 0,
        },
        "ordered_rows_sha256": (
            "a61099b09051bea580be19be7876a65012cb46d6206b8419ce05a76cf8cbadf2"),
    },
    "pdf_vs_tsn_normalized": {
        "side_a": "TSMIS (PDF)", "side_b": "TSN",
        "both_rows": 15_212, "side_a_only_rows": 4,
        "side_b_only_rows": 198, "differing_rows": 783,
        "differing_cells": 1_012,
        "per_field": {
            "PR": 0, "Date of Record": 15, "HG": 364, "Area 4": 58,
            "City Code": 156, "R/U": 68, "Description": 196,
            "Ramp Name": 0, "On/Off": 95, "Ramp Type": 60, "ADT": 0,
        },
        "ordered_rows_sha256": (
            "a61099b09051bea580be19be7876a65012cb46d6206b8419ce05a76cf8cbadf2"),
    },
    "pdf_vs_excel": {
        "side_a": "TSMIS (PDF)", "side_b": "TSMIS (Excel)",
        "both_rows": 15_216, "side_a_only_rows": 0,
        "side_b_only_rows": 0, "differing_rows": 4,
        "differing_cells": 4,
        "per_field": {
            "PR": 0, "Date of Record": 0, "HG": 0, "Area 4": 0,
            "City Code": 0, "R/U": 0, "Description": 4,
            "Ramp Name": 0, "On/Off": 0, "Ramp Type": 0, "ADT": 0,
        },
        "ordered_rows_sha256": (
            "7b08e432e112bd1655eac5a4e6e3dca56d5ed301c2aa445bcfa2a2735d8592c4"),
    },
}


class AuditError(RuntimeError):
    """A source or oracle contract failed closed."""


@dataclass(frozen=True)
class FileEntry:
    name: str
    bytes: int
    sha256: str


@dataclass(frozen=True)
class RampRow:
    source: str
    member: str
    source_record: str
    route: str
    district: str
    county: str
    pr: str
    pm: str
    record_date: str
    hg: str
    area4: str
    city: str
    ru: str
    description: str
    onoff: str = ""
    ramp_type: str = ""
    ramp_name: str = ""
    adt: str = ""
    pm_suffix: str = ""
    adt_eff_year: str = ""
    effective_date: str = ""
    raw_location: str = ""
    raw_area4: str = ""
    raw_description: str = ""
    raw_onoff: str = ""
    raw_ramp_type: str = ""

    @property
    def key(self) -> tuple[str, str, str]:
        return self.route, self.county, self.pm

    @property
    def reference(self) -> str:
        return f"{self.member}:{self.source_record}"


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_bytes(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
        default=str).encode("utf-8")


def _canonical_digest(value: object) -> str:
    return _sha_bytes(_canonical_bytes(value))


def _manifest(root: Path, suffix: str) -> tuple[dict[str, object], list[FileEntry]]:
    paths = sorted(root.glob(f"*{suffix}"), key=lambda path: path.name)
    entries = [FileEntry(path.name, path.stat().st_size, _sha_file(path)) for path in paths]
    payload = "".join(
        f"{entry.name}\t{entry.bytes}\t{entry.sha256}\n" for entry in entries
    ).encode("utf-8")
    return {
        "files": len(entries),
        "bytes": sum(entry.bytes for entry in entries),
        "manifest_sha256": _sha_bytes(payload),
        "serialization": "name\\tbytes\\tsha256\\n sorted by name",
    }, entries


def _require_tree_binding(label: str, root: Path) -> dict[str, object]:
    binding = TREE_BINDINGS[label]
    observed, entries = _manifest(root, str(binding["suffix"]))
    for key in ("files", "bytes", "manifest_sha256"):
        if observed[key] != binding[key]:
            raise AuditError(
                f"{label} {key} drift: {observed[key]!r} != {binding[key]!r}")
    return {"binding": dict(binding), "observed": observed,
            "members": [asdict(entry) for entry in entries]}


def _require_file_binding(label: str, path: Path) -> dict[str, object]:
    observed = {"bytes": path.stat().st_size, "sha256": _sha_file(path)}
    if observed != FILE_BINDINGS[label]:
        raise AuditError(
            f"{label} identity drift: {observed!r} != {FILE_BINDINGS[label]!r}")
    return {"path": str(path.resolve()), "binding": dict(FILE_BINDINGS[label]),
            "observed": observed}


def _file_identity(path: Path) -> dict[str, object]:
    return {"bytes": path.stat().st_size, "sha256": _sha_file(path)}


def _capture_tree(label: str, root: Path, destination: Path) -> tuple[
        dict[str, object], Path]:
    binding = TREE_BINDINGS[label]
    captured_binding = _require_tree_binding(label, root)
    entries = [FileEntry(**entry) for entry in captured_binding["members"]]
    destination.mkdir(parents=True, exist_ok=False)
    for entry in entries:
        payload = (root / entry.name).read_bytes()
        if len(payload) != entry.bytes or _sha_bytes(payload) != entry.sha256:
            raise AuditError(
                f"{label} changed during immutable capture: {entry.name}")
        (destination / entry.name).write_bytes(payload)
    private_manifest, _ = _manifest(destination, str(binding["suffix"]))
    source_after, _ = _manifest(root, str(binding["suffix"]))
    if private_manifest != captured_binding["observed"]:
        raise AuditError(f"{label} private snapshot does not match source manifest")
    if source_after != captured_binding["observed"]:
        raise AuditError(f"{label} changed across immutable capture")
    return captured_binding, destination


def _capture_file(label: str, source: Path, destination: Path) -> tuple[
        dict[str, object], Path]:
    binding = _require_file_binding(label, source)
    payload = source.read_bytes()
    if {"bytes": len(payload), "sha256": _sha_bytes(payload)} != binding["observed"]:
        raise AuditError(f"{label} changed during immutable capture")
    destination.parent.mkdir(parents=True, exist_ok=False)
    destination.write_bytes(payload)
    if destination.read_bytes() != payload:
        raise AuditError(f"{label} private snapshot changed after write")
    if _file_identity(source) != binding["observed"]:
        raise AuditError(f"{label} changed across immutable capture")
    return binding, destination


def _text(value: object, *, upper: bool = False) -> str:
    if value is None:
        out = ""
    elif isinstance(value, bool):
        raise AuditError(f"Boolean is not an admitted scalar: {value!r}")
    elif isinstance(value, str):
        out = value.strip()
    elif isinstance(value, Decimal):
        out = format(value, "f")
    elif isinstance(value, int):
        out = str(value)
    elif isinstance(value, float):
        if not math.isfinite(value):
            raise AuditError(f"non-finite scalar: {value!r}")
        out = str(int(value)) if value.is_integer() else format(Decimal(str(value)), "f")
    else:
        out = str(value).strip()
    return out.upper() if upper else out


def _route(value: object) -> str:
    literal = _text(value, upper=True)
    match = ROUTE_RE.fullmatch(literal)
    if match is None:
        raise AuditError(f"invalid route token: {value!r}")
    return f"{int(match.group(1)):03d}{match.group(2)}"


def _route_from_name(path: Path) -> str:
    matches = ROUTE_NAME_RE.findall(path.stem)
    if len(matches) != 1:
        raise AuditError(f"{path.name}: expected exactly one filename route, got {matches!r}")
    return _route(matches[0])


def _location(value: object) -> tuple[str, str, str, str]:
    literal = _text(value, upper=True)
    match = LOCATION_RE.fullmatch(literal)
    if match is None:
        raise AuditError(f"invalid Ramp Location: {value!r}")
    return (match.group("district"), match.group("county"),
            _route(match.group("route")), literal)


def _pm(value: object) -> str:
    literal = _text(value)
    if not literal or literal.startswith("+"):
        raise AuditError(f"invalid Ramp PM token: {value!r}")
    try:
        parsed = Decimal(literal)
    except InvalidOperation as exc:
        raise AuditError(f"invalid Ramp PM token: {value!r}") from exc
    if not parsed.is_finite() or parsed < 0:
        raise AuditError(f"invalid Ramp PM domain: {value!r}")
    quantized = parsed.quantize(Decimal("0.001"))
    if parsed != quantized:
        raise AuditError(f"Ramp PM has unsupported precision: {value!r}")
    if quantized == 0:
        quantized = Decimal("0.000")
    return format(quantized, ".3f")


def _date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    literal = _text(value)
    for pattern in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(literal, pattern).date().isoformat()
        except ValueError:
            pass
    raise AuditError(f"invalid Ramp date: {value!r}")


def _collapse(value: str) -> str:
    return WS_RE.sub(" ", value).strip()


def _tsmis_description(value: object, route: str, *, pdf: bool) -> tuple[str, str]:
    raw = "" if value is None else str(value)
    literal = raw.strip()
    if not literal or (pdf and literal == NULL_DESCRIPTION):
        return "", raw
    match = TSMIS_DESCRIPTION_RE.fullmatch(literal)
    if match is None:
        return literal, raw
    outer = ROUTE_RE.fullmatch(_text(match.group("route"), upper=True))
    current = ROUTE_RE.fullmatch(route)
    if (outer is not None and current is not None
            and int(outer.group(1)) == int(current.group(1))
            and outer.group(2) in {"", current.group(2)}):
        return match.group("description").strip(), raw
    # A different numeric prefix is source data, not a TSMIS-added outer route.
    return literal, raw


def _row_from_tsmis_values(*, source: str, member: str, source_record: str,
                           file_route: str, values: Sequence[object], pdf: bool) -> RampRow:
    expected = 13 if pdf else 11
    if len(values) != expected:
        raise AuditError(f"{member}:{source_record}: row width {len(values)} != {expected}")
    district, county, route, location = _location(values[0])
    if route != file_route:
        raise AuditError(
            f"{member}:{source_record}: Location route {route} != filename {file_route}")
    description, raw_description = _tsmis_description(values[9], route, pdf=pdf)
    raw_area = _text(values[6], upper=True)
    area = "" if pdf and raw_area == "-" else raw_area
    raw_onoff = _text(values[11], upper=True) if pdf else ""
    onoff = "" if raw_onoff == "-" else ("O" if raw_onoff == "N" else raw_onoff)
    raw_rtype = _text(values[12], upper=True) if pdf else ""
    if not pdf and (values[4] is not None or values[10] is not None):
        raise AuditError(
            f"{member}:{source_record}: declared blank TSMIS cells carry data")
    if pdf and (values[4] not in (None, "") or values[10] not in (None, "")):
        raise AuditError(
            f"{member}:{source_record}: declared blank PDF projection cells carry data")
    return RampRow(
        source=source, member=member, source_record=source_record,
        route=route, district=district, county=county,
        pr=_text(values[1], upper=True), pm=_pm(values[2]),
        record_date=_date(values[3]), hg=_text(values[5], upper=True),
        area4=area, city=_text(values[7], upper=True),
        ru=_text(values[8], upper=True), description=description,
        onoff=onoff, ramp_type=raw_rtype,
        raw_location=location, raw_area4=raw_area,
        raw_description=raw_description, raw_onoff=raw_onoff,
        raw_ramp_type=raw_rtype)


def _parse_tsmis_xlsx(root: Path) -> dict[str, object]:
    rows: list[RampRow] = []
    routes: list[str] = []
    route_counts: dict[str, int] = {}
    formula_cells = []
    error_cells = []
    for path in sorted(root.glob("*.xlsx"), key=lambda item: item.name):
        file_route = _route_from_name(path)
        if file_route in route_counts:
            raise AuditError(f"duplicate TSMIS XLSX route: {file_route}")
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if workbook.sheetnames != [TSMIS_SHEET]:
                raise AuditError(f"{path.name}: sheet universe {workbook.sheetnames!r}")
            worksheet = workbook[TSMIS_SHEET]
            physical = worksheet.iter_rows()
            header_cells = next(physical, ())
            header = tuple(cell.value for cell in header_cells)
            if header != TSMIS_HEADER or worksheet.max_column != len(TSMIS_HEADER):
                raise AuditError(f"{path.name}: exact header/layout drift: {header!r}")
            before = len(rows)
            for source_row, cells in enumerate(physical, 2):
                values = tuple(cell.value for cell in cells)
                if all(value is None for value in values):
                    raise AuditError(f"{path.name}: blank physical row {source_row}")
                for cell in cells:
                    if cell.data_type == "f":
                        formula_cells.append((path.name, cell.coordinate))
                    elif cell.data_type == "e":
                        error_cells.append((path.name, cell.coordinate, cell.value))
                rows.append(_row_from_tsmis_values(
                    source="TSMIS Excel", member=path.name,
                    source_record=f"row {source_row}", file_route=file_route,
                    values=values, pdf=False))
            route_counts[file_route] = len(rows) - before
            routes.append(file_route)
        finally:
            workbook.close()
    if formula_cells or error_cells:
        raise AuditError(
            f"TSMIS XLSX formula/error cells: {formula_cells[:5]!r} {error_cells[:5]!r}")
    return _source_summary(rows, routes, route_counts, extra={
        "formula_cells": 0, "error_cells": 0,
        "openpyxl_version": openpyxl.__version__,
    })


def _pdf_header(words: Sequence[dict[str, object]]) -> tuple[dict[str, float], float] | None:
    token_words = [word for word in words if str(word["text"]) in PDF_HEADER_TOKENS]
    top_groups: list[list[dict[str, object]]] = []
    for word in sorted(token_words, key=lambda item: (float(item["top"]), float(item["x0"]))):
        if top_groups and abs(float(word["top"]) - float(top_groups[-1][0]["top"])) <= 2.0:
            top_groups[-1].append(word)
        else:
            top_groups.append([word])
    candidates = [
        group for group in top_groups
        if set(str(word["text"]) for word in group) == set(PDF_HEADER_TOKENS)
    ]
    if not candidates:
        return None
    if len(candidates) != 1:
        raise AuditError(f"Ramp PDF page has {len(candidates)} complete header constellations")
    header_words = candidates[0]
    anchors: dict[str, dict[str, object]] = {}
    for token in PDF_HEADER_TOKENS:
        matches = [word for word in header_words if str(word["text"]) == token]
        if len(matches) != 1:
            raise AuditError(f"Ramp PDF header token {token!r} occurs {len(matches)} times")
        anchors[token] = matches[0]
    top = float(anchors["DESCRIPTION"]["top"])
    singles = [word for word in words
               if abs(float(word["top"]) - top) <= 2.0
               and len(str(word["text"])) == 1]
    prefix = [word for word in singles
              if float(anchors["LOCATION"]["x1"]) < float(word["x0"])
              < float(anchors["PM"]["x0"])]
    tail = sorted(
        (word for word in singles
         if float(anchors["CODE"]["x1"]) < float(word["x0"])
         < float(anchors["DESCRIPTION"]["x0"])),
        key=lambda word: float(word["x0"]))
    if len(prefix) != 1 or len(tail) != 3:
        return None
    ru, onoff, rtype = tail
    boundaries = {
        "loc_pr": float(prefix[0]["x0"]) - 6.0,
        "pr_pm": float(anchors["PM"]["x0"]) - 12.0,
        "pm_date": float(anchors["RECORD"]["x0"]) - 8.0,
        "date_hg": float(anchors["RECORD"]["x1"]) + 12.0,
        "hg_area": float(anchors["AREA"]["x0"]) - 8.0,
        "area_city": float(anchors["CITY"]["x0"]) - 8.0,
        "city_ru": float(ru["x0"]) - 6.0,
        "ru_onoff": (float(ru["x1"]) + float(onoff["x0"])) / 2.0,
        "onoff_type": (float(onoff["x1"]) + float(rtype["x0"])) / 2.0,
        "type_desc": float(anchors["DESCRIPTION"]["x0"]) - 8.0,
    }
    if list(boundaries.values()) != sorted(boundaries.values()):
        raise AuditError(f"Ramp PDF column boundaries are not strictly ordered: {boundaries}")
    bottom = max(float(word["bottom"]) for word in [*anchors.values(), *singles])
    return boundaries, bottom


def _cluster_pdf_lines(words: Sequence[dict[str, object]]) -> list[tuple[float, list[dict[str, object]]]]:
    lines: list[list[object]] = []
    for word in sorted(words, key=lambda item: (float(item["top"]), float(item["x0"]))):
        top = float(word["top"])
        if lines and abs(top - float(lines[-1][0])) <= 2.0:
            lines[-1][1].append(word)  # type: ignore[union-attr]
        else:
            lines.append([top, [word]])
    return [(float(top), sorted(group, key=lambda item: float(item["x0"])))
            for top, group in lines]  # type: ignore[arg-type]


PDF_COLUMN_ORDER = (
    "location", "pr", "pm", "date", "hg", "area4", "city", "ru",
    "onoff", "ramp_type", "description")


def _classify_pdf_line(words: Sequence[dict[str, object]],
                       boundaries: dict[str, float]) -> dict[str, str]:
    values: dict[str, list[str]] = {name: [] for name in PDF_COLUMN_ORDER}
    cuts = (
        ("description", boundaries["type_desc"]),
        ("ramp_type", boundaries["onoff_type"]),
        ("onoff", boundaries["ru_onoff"]),
        ("ru", boundaries["city_ru"]),
        ("city", boundaries["area_city"]),
        ("area4", boundaries["hg_area"]),
        ("hg", boundaries["date_hg"]),
        ("date", boundaries["pm_date"]),
        ("pm", boundaries["pr_pm"]),
        ("pr", boundaries["loc_pr"]),
    )
    for word in words:
        center = (float(word["x0"]) + float(word["x1"])) / 2.0
        target = "location"
        for name, cut in cuts:
            if center >= cut:
                target = name
                break
        values[target].append(str(word["text"]))
    return {name: " ".join(parts) for name, parts in values.items()}


def _parse_tsmis_pdf(root: Path) -> dict[str, object]:
    rows: list[RampRow] = []
    routes: list[str] = []
    route_counts: dict[str, int] = {}
    page_counts: dict[str, int] = {}
    data_pages = 0
    unclassified = []
    fragments = []
    unexpected_prefixes = []
    header_geometry = []
    for path in sorted(root.glob("*.pdf"), key=lambda item: item.name):
        file_route = _route_from_name(path)
        if file_route in route_counts:
            raise AuditError(f"duplicate TSMIS PDF route: {file_route}")
        reader = PdfReader(str(path))
        if len(reader.pages) < 2:
            raise AuditError(f"{path.name}: PDF has no data page")
        cover = reader.pages[0].extract_text() or ""
        compact_cover = re.sub(r"\s+", " ", cover)
        if file_route not in compact_cover or "REPORT TITLE" not in compact_cover:
            raise AuditError(f"{path.name}: cover provenance/route drift")
        before = len(rows)
        member_data_pages = 0
        with pdfplumber.open(path) as document:
            if len(document.pages) != len(reader.pages):
                raise AuditError(f"{path.name}: PDF reader page-count disagreement")
            for page_number, page in enumerate(document.pages, 1):
                words = page.extract_words()
                header = _pdf_header(words)
                if header is None:
                    if page_number != 1:
                        raise AuditError(
                            f"{path.name}: non-cover page {page_number} lacks exact table header")
                    continue
                if page_number == 1:
                    raise AuditError(f"{path.name}: cover page unexpectedly has data header")
                member_data_pages += 1
                data_pages += 1
                boundaries, header_bottom = header
                header_geometry.append((path.name, page_number, boundaries, header_bottom))
                page_rows: list[tuple[float, list[object], int]] = []
                page_fragments: list[tuple[float, str]] = []
                for top, line_words in _cluster_pdf_lines(words):
                    if top <= header_bottom + 2.0:
                        continue
                    values = _classify_pdf_line(line_words, boundaries)
                    if PDF_PM_RE.fullmatch(values["pm"]):
                        if values["pr"] and values["pr"] not in PDF_PREFIXES:
                            unexpected_prefixes.append(
                                (path.name, page_number, values["pr"], values["pm"]))
                        projected: list[object] = [
                            values["location"], values["pr"], values["pm"], values["date"],
                            None, values["hg"], values["area4"], values["city"],
                            values["ru"], values["description"], None,
                            values["onoff"], values["ramp_type"],
                        ]
                        page_rows.append((top, projected, len(page_rows) + 1))
                    elif values["description"] and not any(
                            values[name] for name in PDF_COLUMN_ORDER
                            if name != "description"):
                        page_fragments.append((top, values["description"]))
                    else:
                        unclassified.append({
                            "member": path.name, "page": page_number, "top": top,
                            "text": " ".join(str(word["text"]) for word in line_words),
                            "columns": values,
                        })
                description_parts = {
                    index: [(top, str(projected[9]))]
                    for index, (top, projected, _row_on_page) in enumerate(page_rows)}
                for top, text in page_fragments:
                    if not page_rows:
                        fragments.append((path.name, page_number, top, text, "no row"))
                        continue
                    distances = [abs(top - row_top) for row_top, _row, _ordinal in page_rows]
                    nearest = min(range(len(distances)), key=distances.__getitem__)
                    if distances[nearest] > 13.0:
                        fragments.append(
                            (path.name, page_number, top, text, distances[nearest]))
                        continue
                    description_parts[nearest].append((top, text))
                for index, (_top, projected, row_on_page) in enumerate(page_rows):
                    combined = ""
                    for _fragment_top, text in sorted(description_parts[index]):
                        if not text:
                            continue
                        if not combined or combined.endswith("-"):
                            combined += text
                        else:
                            combined += " " + text
                    projected[9] = combined
                    rows.append(_row_from_tsmis_values(
                        source="TSMIS PDF", member=path.name,
                        source_record=f"page {page_number} row {row_on_page}",
                        file_route=file_route, values=tuple(projected), pdf=True))
        if member_data_pages != len(reader.pages) - 1:
            raise AuditError(
                f"{path.name}: data-page count {member_data_pages} != pages-1")
        route_counts[file_route] = len(rows) - before
        page_counts[file_route] = len(reader.pages)
        routes.append(file_route)
    if unclassified or fragments or unexpected_prefixes:
        raise AuditError(
            "TSMIS PDF parse residue: "
            f"unclassified={unclassified[:3]!r} fragments={fragments[:3]!r} "
            f"prefixes={unexpected_prefixes[:3]!r}")
    return _source_summary(rows, routes, route_counts, extra={
        "pages": sum(page_counts.values()),
        "data_pages": data_pages,
        "page_counts": page_counts,
        "unclassified_lines": 0,
        "unattached_description_fragments": 0,
        "unexpected_postmile_prefixes": 0,
        "header_geometry_sha256": _canonical_digest(header_geometry),
    })


def _parse_tsn_xlsx(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    rows: list[RampRow] = []
    formula_cells = []
    error_cells = []
    try:
        if workbook.sheetnames != [TSN_SHEET]:
            raise AuditError(f"TSN sheet universe {workbook.sheetnames!r}")
        worksheet = workbook[TSN_SHEET]
        physical = worksheet.iter_rows()
        header_cells = next(physical, ())
        header = tuple(cell.value for cell in header_cells)
        if header != TSN_HEADER or worksheet.max_column != len(TSN_HEADER):
            raise AuditError(f"TSN exact header/layout drift: {header!r}")
        index = {name: offset for offset, name in enumerate(TSN_HEADER)}
        for source_row, cells in enumerate(physical, 2):
            values = tuple(cell.value for cell in cells)
            if all(value is None for value in values):
                raise AuditError(f"TSN blank physical row {source_row}")
            for cell in cells:
                if cell.data_type == "f":
                    formula_cells.append(cell.coordinate)
                elif cell.data_type == "e":
                    error_cells.append((cell.coordinate, cell.value))
            get = lambda name: values[index[name]]
            district, county, route, location = _location(get("LOCATION"))
            description = _text(get("DESCRIPTION"))
            rows.append(RampRow(
                source="TSN XLSX", member=path.name,
                source_record=f"row {source_row}", route=route,
                district=district, county=county,
                pr=_text(get("PR"), upper=True), pm=_pm(get("PM")),
                record_date=_date(get("DATE_OF_RECORD")),
                hg=_text(get("HG"), upper=True),
                area4=_text(get("AREA_4"), upper=True),
                city=_text(get("CITY_CODE"), upper=True),
                ru=_text(get("POP"), upper=True), description=description,
                onoff=_text(get("ON_OFF"), upper=True),
                ramp_type=_text(get("RAMP_TYPE"), upper=True),
                ramp_name=_text(get("RAMP_NANE")), adt=_text(get("ADT")),
                pm_suffix=_text(get("PM_SFX"), upper=True),
                adt_eff_year=_text(get("ADT_EFF_YEAR")),
                effective_date=_date(get("EFF_DATE")), raw_location=location,
                raw_area4=_text(get("AREA_4"), upper=True),
                raw_description="" if get("DESCRIPTION") is None else str(get("DESCRIPTION")),
                raw_onoff=_text(get("ON_OFF"), upper=True),
                raw_ramp_type=_text(get("RAMP_TYPE"), upper=True)))
    finally:
        workbook.close()
    if formula_cells or error_cells:
        raise AuditError(
            f"TSN formula/error cells: {formula_cells[:5]!r} {error_cells[:5]!r}")
    route_counts = Counter(row.route for row in rows)
    routes = list(dict.fromkeys(row.route for row in rows))
    return _source_summary(rows, routes, dict(route_counts), extra={
        "formula_cells": 0, "error_cells": 0,
        "openpyxl_version": openpyxl.__version__,
    })


def _parse_tsn_normalized(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    rows: list[RampRow] = []
    formula_cells = []
    error_cells = []
    try:
        if workbook.sheetnames != [TSN_NORMALIZED_SHEET]:
            raise AuditError(
                f"normalized TSN sheet universe {workbook.sheetnames!r}")
        worksheet = workbook[TSN_NORMALIZED_SHEET]
        physical = worksheet.iter_rows()
        header_cells = next(physical, ())
        header = tuple(cell.value for cell in header_cells)
        # This accepted workbook intentionally has no cached worksheet dimension;
        # openpyxl therefore reports max_column=None until it scans the stream.
        # Gate the physical header and every physical data row instead of trusting
        # optional dimension metadata.
        if header != TSN_NORMALIZED_HEADER:
            raise AuditError(f"normalized TSN exact header/layout drift: {header!r}")
        index = {name: offset for offset, name in enumerate(TSN_NORMALIZED_HEADER)}
        for source_row, cells in enumerate(physical, 2):
            values = tuple(cell.value for cell in cells)
            if len(values) != len(TSN_NORMALIZED_HEADER):
                raise AuditError(
                    f"normalized TSN row {source_row}: physical width "
                    f"{len(values)} != {len(TSN_NORMALIZED_HEADER)}")
            if all(value is None for value in values):
                raise AuditError(f"normalized TSN blank physical row {source_row}")
            for cell in cells:
                if cell.data_type == "f":
                    formula_cells.append(cell.coordinate)
                elif cell.data_type == "e":
                    error_cells.append((cell.coordinate, cell.value))
            get = lambda name: values[index[name]]
            route = _route(get("Route"))
            district = _text(get("TSN District"), upper=True).zfill(2)
            county = _text(get("TSN County"), upper=True)
            if not re.fullmatch(r"\d{2}", district) or not re.fullmatch(
                    r"[A-Z]{2,3}", county):
                raise AuditError(
                    f"normalized TSN row {source_row}: invalid D/C {district}/{county}")
            description = _text(get("Description"))
            rows.append(RampRow(
                source="TSN normalized r7", member=path.name,
                source_record=f"row {source_row}", route=route,
                district=district, county=county,
                pr=_text(get("PR"), upper=True), pm=_pm(get("PM")),
                record_date=_date(get("Date of Record")),
                hg=_text(get("HG"), upper=True),
                area4=_text(get("Area 4"), upper=True),
                city=_text(get("City Code"), upper=True),
                ru=_text(get("R/U"), upper=True), description=description,
                onoff=_text(get("On/Off"), upper=True),
                ramp_type=_text(get("Ramp Type"), upper=True),
                ramp_name=_text(get("Ramp Name")), adt=_text(get("ADT")),
                raw_location=f"{district}-{county}-{route}",
                raw_area4=_text(get("Area 4"), upper=True),
                raw_description="" if get("Description") is None else str(get("Description")),
                raw_onoff=_text(get("On/Off"), upper=True),
                raw_ramp_type=_text(get("Ramp Type"), upper=True)))
    finally:
        workbook.close()
    if formula_cells or error_cells:
        raise AuditError(
            f"normalized TSN formula/error cells: {formula_cells[:5]!r} "
            f"{error_cells[:5]!r}")
    route_counts = Counter(row.route for row in rows)
    routes = list(dict.fromkeys(row.route for row in rows))
    return _source_summary(rows, routes, dict(route_counts), extra={
        "formula_cells": 0, "error_cells": 0,
        "openpyxl_version": openpyxl.__version__,
    })


def _load_normalized_sidecar(path: Path) -> dict[str, object]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    manifest = payload.get("tsn_raw_manifest", {})
    members = manifest.get("members", [])
    normalized = payload.get("tsn_normalized_workbook_identity", {})
    checks = {
        "schema_version_1": payload.get("schema_version") == 1,
        "completion_exact": (
            payload.get("completion") == "complete"
            and payload.get("skipped_inputs") == 0
            and payload.get("failed_inputs") == 0),
        "normalization_version_3": payload.get("tsn_normalization_version") == 3,
        "raw_manifest_exact": (
            manifest.get("version") == 1
            and manifest.get("algorithm") == "sha256"
            and manifest.get("member_count") == 1
            and manifest.get("byte_length") == FILE_BINDINGS["tsn_xlsx"]["bytes"]
            and members == [{
                "relative_path": DEFAULT_TSN_XLSX.name,
                "byte_length": FILE_BINDINGS["tsn_xlsx"]["bytes"],
                "sha256": FILE_BINDINGS["tsn_xlsx"]["sha256"],
            }]),
        "normalized_identity_exact": (
            normalized.get("version") == 1
            and normalized.get("algorithm") == "sha256"
            and normalized.get("byte_length")
            == FILE_BINDINGS["tsn_normalized"]["bytes"]
            and normalized.get("sha256")
            == FILE_BINDINGS["tsn_normalized"]["sha256"]),
        "artifact_token_exact": payload.get("tsn_artifact_identity_token") == (
            "tsn-normalized-v1:"
            "89326764f78c1a1bd1027d49d34eb6b0dd8c18ce868d3e6327301e407c1d2d8c"),
    }
    if not all(checks.values()):
        raise AuditError(f"normalized Ramp Detail sidecar drift: {checks!r}")
    return {
        "checks": checks,
        "raw_manifest": manifest,
        "normalized_workbook_identity": normalized,
        "artifact_identity_token": payload.get("tsn_artifact_identity_token"),
    }


def _load_stage6(result_path: Path, acceptance_path: Path) -> dict[str, object]:
    result = json.loads(result_path.read_text(encoding="utf-8"))
    acceptance = json.loads(acceptance_path.read_text(encoding="utf-8"))
    residue = result.get("classified_projection_residue", {}).get(
        "description_prefix_loss", {})
    invariants = result.get("audit_invariants", {})
    blocking = result.get("findings", {}).get("blocking", [])
    checks = {
        "schema_version_3": result.get("schema_version") == 3,
        "audit_complete": result.get("stage6_family_audit_complete") is True,
        "projection_red": result.get("projection_exact") is False,
        "full_conservation_red": result.get("normalized_full_conservation") is False,
        "all_invariants_true": bool(invariants) and all(invariants.values()),
        "description_contract_exact_15": (
            residue.get("exact") is True
            and residue.get("expected_count") == 15
            and residue.get("observed_count") == 15),
        "blocking_ids_exact": [item.get("id") for item in blocking]
        == ["RD-S6-001", "RD-S6-002", "RD-S6-003"],
        "acceptance_result_identity_exact": (
            acceptance.get("result_bytes") == FILE_BINDINGS["stage6_result"]["bytes"]
            and acceptance.get("result_sha256")
            == FILE_BINDINGS["stage6_result"]["sha256"]),
        "acceptance_audit_complete": (
            acceptance.get("stage6_family_audit_complete") is True),
        "acceptance_postwrite_revalidation": (
            acceptance.get("post_result_write_revalidation") is True),
    }
    if not all(checks.values()):
        raise AuditError(f"accepted Ramp Stage-6 dependency drift: {checks!r}")
    return {
        "checks": checks,
        "description_prefix_loss": residue,
        "blocking_findings": blocking,
        "audit_invariants": invariants,
        "result_sha256": FILE_BINDINGS["stage6_result"]["sha256"],
        "acceptance_sha256": FILE_BINDINGS["stage6_acceptance"]["sha256"],
    }


def _load_tsn_cross_format(path: Path) -> dict[str, object]:
    result = json.loads(path.read_text(encoding="utf-8"))
    sources = result.get("sources", {})
    xlsx = result.get("xlsx", {})
    pdf = result.get("detail_pdf", {})
    parity = result.get("detail_cross_format", {})
    classification = result.get("classification", {})
    checks = {
        "oracle_version_3": result.get("oracle_version") == 3,
        "status_ok": result.get("status") == "ok",
        "xlsx_identity_exact": (
            sources.get("xlsx", {}).get("bytes") == FILE_BINDINGS["tsn_xlsx"]["bytes"]
            and sources.get("xlsx", {}).get("sha256")
            == FILE_BINDINGS["tsn_xlsx"]["sha256"]
            and sources.get("xlsx", {}).get("binding_ok") is True),
        "pdf_identity_exact": (
            sources.get("detail_pdf", {}).get("bytes") == FILE_BINDINGS["tsn_pdf"]["bytes"]
            and sources.get("detail_pdf", {}).get("sha256")
            == FILE_BINDINGS["tsn_pdf"]["sha256"]
            and sources.get("detail_pdf", {}).get("binding_ok") is True),
        "xlsx_rows_exact": xlsx.get("data_rows") == 15_410,
        "pdf_rows_pages_exact": (
            pdf.get("parsed_rows") == 15_410
            and pdf.get("physical_pages") == 500
            and pdf.get("data_pages") == 498),
        "cross_format_exact": (
            parity.get("exact_render_equivalence") == 15_410
            and parity.get("exact_extracted_content") == 15_404
            and parity.get("proven_print_truncation_equivalence_count") == 6
            and parity.get("matched_keys") == 15_410
            and parity.get("unproven_differences") == []),
        "zero_unresolved": (
            result.get("unresolved_residue") == []
            and classification.get("unresolved_residue_count") == 0),
        "all_18_fields_dispositioned": xlsx.get("field_disposition", {}).get(
            "all_18_columns_accounted_once") is True,
        "negative_self_check_pass": result.get(
            "internal_negative_mutation_self_check", {}).get("pass") is True,
    }
    if not all(checks.values()):
        raise AuditError(f"accepted Ramp TSN cross-format dependency drift: {checks!r}")
    return {
        "checks": checks,
        "parity": result.get("parity"),
        "classification": classification,
        "six_print_clip_equivalences": parity.get(
            "proven_print_truncation_equivalences"),
        "four_text_extraction_artifacts": pdf.get(
            "pdf_text_extraction_artifacts"),
        "xlsx_field_disposition": xlsx.get("field_disposition"),
        "source_dates": result.get("source_dates"),
        "result_sha256": FILE_BINDINGS["tsn_cross_format"]["sha256"],
    }


def _source_summary(rows: Sequence[RampRow], routes: Sequence[str],
                    route_counts: dict[str, int], *, extra: dict[str, object]) -> dict[str, object]:
    keys = Counter(row.key for row in rows)
    duplicates = [
        {"key": list(key), "occurrences": count,
         "references": [row.reference for row in rows if row.key == key]}
        for key, count in sorted(keys.items()) if count > 1]
    return {
        "rows_data": list(rows),
        "rows": len(rows),
        "routes": list(routes),
        "route_count": len(routes),
        "route_counts": dict(route_counts),
        "identity": {
            "policy": ["Route", "County", "norm_pm(PM)"],
            "unique_keys": len(keys),
            "duplicate_groups": len(duplicates),
            "duplicate_occurrences_beyond_first": sum(
                item["occurrences"] - 1 for item in duplicates),
            "duplicates": duplicates,
        },
        "ordered_row_sha256": _canonical_digest([
            {key: value for key, value in asdict(row).items()
             if not key.startswith("raw_")}
            for row in rows]),
        "raw_representation_sha256": _canonical_digest([asdict(row) for row in rows]),
        **extra,
    }


def _weak_identity_census(rows: Sequence[RampRow]) -> dict[str, object]:
    groups: dict[tuple[str, str], list[RampRow]] = defaultdict(list)
    for row in rows:
        groups[(row.route, row.pm)].append(row)
    cross_county = []
    for weak_key, members in sorted(groups.items()):
        counties = sorted({row.county for row in members})
        if len(counties) < 2:
            continue
        identities = sorted({row.key for row in members})
        cross_county.append({
            "weak_key": list(weak_key),
            "counties": counties,
            "county_identity_count": len(identities),
            "row_count": len(members),
            "identities": [list(key) for key in identities],
            "references": [row.reference for row in sorted(
                members, key=lambda item: item.reference)],
        })
    return {
        "weak_identity": ["Route", "norm_pm(PM)"],
        "approved_identity": ["Route", "County", "norm_pm(PM)"],
        "weak_key_count": len(groups),
        "cross_county_weak_key_count": len(cross_county),
        "cross_county_identity_count": sum(
            item["county_identity_count"] for item in cross_county),
        "cross_county_row_count": sum(item["row_count"] for item in cross_county),
        "groups": cross_county,
        "ordered_sha256": _canonical_digest(cross_county),
    }


def _weak_identity_semantics(census: dict[str, object]) -> list[dict[str, object]]:
    return [
        {key: group[key] for key in (
            "weak_key", "counties", "county_identity_count", "row_count",
            "identities")}
        for group in census["groups"]
    ]


def _source_claim_contract(tsn_rows: Sequence[RampRow]) -> dict[str, object]:
    suffix_rows = [row for row in tsn_rows if row.pm_suffix]
    suffix_counts = Counter(row.pm_suffix for row in suffix_rows)
    suffix_mismatches = [
        {"reference": row.reference, "key": list(row.key),
         "pm_suffix": row.pm_suffix, "hg": row.hg}
        for row in suffix_rows if row.pm_suffix != row.hg]
    blank_suffix_hg_lr = [
        {"reference": row.reference, "key": list(row.key), "hg": row.hg}
        for row in tsn_rows if not row.pm_suffix and row.hg in {"L", "R"}]
    return {
        "pm_suffix_nonblank": len(suffix_rows),
        "pm_suffix_counts": dict(sorted(suffix_counts.items())),
        "all_nonblank_pm_suffix_equals_hg": not suffix_mismatches,
        "nonblank_suffix_hg_mismatches": suffix_mismatches,
        "blank_suffix_with_hg_l_or_r": blank_suffix_hg_lr,
        "adt_effective_year_nonblank": sum(
            bool(row.adt_eff_year) for row in tsn_rows),
        "effective_date_nonblank": sum(
            bool(row.effective_date) for row in tsn_rows),
        "ramp_name_nonblank": sum(bool(row.ramp_name) for row in tsn_rows),
        "adt_nonblank": sum(bool(row.adt) for row in tsn_rows),
    }


def _comparison_row(row: RampRow, fields: Sequence[str], *, collapse_description: bool) -> tuple[str, ...]:
    values = []
    for field in fields:
        value = getattr(row, field)
        if field == "description" and collapse_description:
            value = _collapse(value)
        values.append(value)
    return tuple(values)


def _pair_group(left: Sequence[RampRow], right: Sequence[RampRow],
                fields: Sequence[str], *, collapse_description: bool) -> tuple[
                    list[tuple[RampRow, RampRow]], list[RampRow], list[RampRow], dict[str, object]]:
    left = sorted(left, key=lambda row: row.reference)
    right = sorted(right, key=lambda row: row.reference)
    if not left or not right:
        return [], list(left), list(right), {"candidate_assignments": 0, "cost": 0}
    if max(len(left), len(right)) > 8:
        raise AuditError(
            f"Ramp duplicate group exceeds independent exact cap: {len(left)}x{len(right)}")
    costs = [[
        sum(a != b for a, b in zip(
            _comparison_row(lrow, fields, collapse_description=collapse_description),
            _comparison_row(rrow, fields, collapse_description=collapse_description)))
        for rrow in right] for lrow in left]
    candidate_count = 0
    best: tuple[object, ...] | None = None
    best_pairs: list[tuple[int, int]] = []
    if len(left) <= len(right):
        for chosen in itertools.permutations(range(len(right)), len(left)):
            candidate_count += 1
            pairs = [(left_index, right_index)
                     for left_index, right_index in enumerate(chosen)]
            rank = (sum(costs[i][j] for i, j in pairs), chosen)
            if best is None or rank < best:
                best, best_pairs = rank, pairs
    else:
        for chosen in itertools.permutations(range(len(left)), len(right)):
            candidate_count += 1
            pairs = [(left_index, right_index)
                     for right_index, left_index in enumerate(chosen)]
            rank = (sum(costs[i][j] for i, j in pairs), chosen)
            if best is None or rank < best:
                best, best_pairs = rank, pairs
    selected_left = {i for i, _j in best_pairs}
    selected_right = {j for _i, j in best_pairs}
    paired = [(left[i], right[j]) for i, j in sorted(best_pairs)]
    return (
        paired,
        [row for index, row in enumerate(left) if index not in selected_left],
        [row for index, row in enumerate(right) if index not in selected_right],
        {"candidate_assignments": candidate_count,
         "cost": int(best[0]) if best is not None else 0,
         "cost_matrix": costs,
         "selected": [[i, j] for i, j in sorted(best_pairs)]},
    )


def _compare(label: str, left_rows: Sequence[RampRow], right_rows: Sequence[RampRow],
             fields: Sequence[str], *, collapse_description: bool,
             context_fields: Sequence[str] = ()) -> dict[str, object]:
    left_by_key: dict[tuple[str, str, str], list[RampRow]] = defaultdict(list)
    right_by_key: dict[tuple[str, str, str], list[RampRow]] = defaultdict(list)
    for row in left_rows:
        left_by_key[row.key].append(row)
    for row in right_rows:
        right_by_key[row.key].append(row)
    paired_count = differing_rows = differing_cells = 0
    per_field: Counter[str] = Counter()
    only_left: list[dict[str, object]] = []
    only_right: list[dict[str, object]] = []
    differing: list[dict[str, object]] = []
    pair_wire = []
    duplicate_traces = []
    for key in sorted(set(left_by_key) | set(right_by_key)):
        paired, left_extra, right_extra, trace = _pair_group(
            left_by_key[key], right_by_key[key], fields,
            collapse_description=collapse_description)
        if len(left_by_key[key]) > 1 or len(right_by_key[key]) > 1:
            duplicate_traces.append({
                "key": list(key),
                "left": [row.reference for row in sorted(
                    left_by_key[key], key=lambda item: item.reference)],
                "right": [row.reference for row in sorted(
                    right_by_key[key], key=lambda item: item.reference)],
                **trace,
            })
        for row in left_extra:
            only_left.append({"key": list(key), "reference": row.reference,
                              "values": asdict(row)})
        for row in right_extra:
            only_right.append({"key": list(key), "reference": row.reference,
                               "values": asdict(row)})
        for left, right in paired:
            paired_count += 1
            left_values = _comparison_row(
                left, fields, collapse_description=collapse_description)
            right_values = _comparison_row(
                right, fields, collapse_description=collapse_description)
            differences = []
            for field, left_value, right_value in zip(fields, left_values, right_values):
                if left_value != right_value:
                    differing_cells += 1
                    per_field[field] += 1
                    differences.append({
                        "field": field, "left": left_value, "right": right_value})
            pair_wire.append({
                "key": list(key), "left": left.reference, "right": right.reference,
                "left_values": list(left_values), "right_values": list(right_values)})
            if differences:
                differing_rows += 1
                differing.append({
                    "key": list(key), "left": left.reference,
                    "right": right.reference, "differences": differences})
    context = {}
    for field in context_fields:
        left_values = [getattr(row, field) for row in left_rows]
        right_values = [getattr(row, field) for row in right_rows]
        context[field] = {
            "left_nonblank": sum(bool(value) for value in left_values),
            "right_nonblank": sum(bool(value) for value in right_values),
            "left_typed_sha256": _canonical_digest(left_values),
            "right_typed_sha256": _canonical_digest(right_values),
            "asserted": False,
        }
    return {
        "label": label,
        "identity_policy": ["Route", "County", "norm_pm(PM)"],
        "asserted_fields": list(fields),
        "context_fields": context,
        "collapse_description_whitespace": collapse_description,
        "counts": {
            "left_rows": len(left_rows), "right_rows": len(right_rows),
            "paired_rows": paired_count,
            "left_only_rows": len(only_left),
            "right_only_rows": len(only_right),
            "differing_rows": differing_rows,
            "identical_rows": paired_count - differing_rows,
            "differing_cells": differing_cells,
            "asserted_cells": paired_count * len(fields),
            "per_field": {field: per_field.get(field, 0) for field in fields},
        },
        "paired_ledger_sha256": _canonical_digest(pair_wire),
        "differing_rows": differing,
        "left_only": only_left,
        "right_only": only_right,
        "duplicate_pairing": {
            "policy": "exhaustive minimum asserted-cell cost; lexical source-order tie break",
            "groups": duplicate_traces,
            "group_count": len(duplicate_traces),
        },
    }


def _paired_rows(left_rows: Sequence[RampRow], right_rows: Sequence[RampRow],
                 fields: Sequence[str], *, collapse_description: bool) -> tuple[
                     list[tuple[RampRow, RampRow]], list[RampRow], list[RampRow]]:
    left_by_key: dict[tuple[str, str, str], list[RampRow]] = defaultdict(list)
    right_by_key: dict[tuple[str, str, str], list[RampRow]] = defaultdict(list)
    for row in left_rows:
        left_by_key[row.key].append(row)
    for row in right_rows:
        right_by_key[row.key].append(row)
    pairs: list[tuple[RampRow, RampRow]] = []
    left_only: list[RampRow] = []
    right_only: list[RampRow] = []
    for key in sorted(set(left_by_key) | set(right_by_key)):
        selected, extra_left, extra_right, _trace = _pair_group(
            left_by_key[key], right_by_key[key], fields,
            collapse_description=collapse_description)
        pairs.extend(selected)
        left_only.extend(extra_left)
        right_only.extend(extra_right)
    return pairs, left_only, right_only


def _description_source_contract(excel_rows: Sequence[RampRow],
                                 pdf_rows: Sequence[RampRow],
                                 tsn_rows: Sequence[RampRow]) -> dict[str, object]:
    excel_pairs, _excel_only, _tsn_only = _paired_rows(
        excel_rows, tsn_rows, BASE_ASSERTED_FIELDS,
        collapse_description=False)
    pdf_pairs, _pdf_only, _tsn_only_pdf = _paired_rows(
        pdf_rows, tsn_rows, PDF_TSN_ASSERTED_FIELDS,
        collapse_description=True)
    excel_by_tsn = {right.reference: left for left, right in excel_pairs}
    pdf_by_tsn = {right.reference: left for left, right in pdf_pairs}
    prefixed = []
    for tsn in tsn_rows:
        match = re.match(r"^\d+/", tsn.description)
        if match is None:
            continue
        excel = excel_by_tsn.get(tsn.reference)
        pdf = pdf_by_tsn.get(tsn.reference)
        prefixed.append({
            "key": list(tsn.key),
            "tsn_reference": tsn.reference,
            "tsn_description": tsn.description,
            "excel_reference": excel.reference if excel else None,
            "excel_description": excel.description if excel else None,
            "excel_raw_description": excel.raw_description if excel else None,
            "pdf_reference": pdf.reference if pdf else None,
            "pdf_description": pdf.description if pdf else None,
            "pdf_raw_description": pdf.raw_description if pdf else None,
            "excel_exact": bool(excel and excel.description == tsn.description),
            "pdf_render_exact": bool(
                pdf and _collapse(pdf.description) == _collapse(tsn.description)),
        })

    outer_census: Counter[str] = Counter()
    suffix_examples = []
    for row in excel_rows:
        literal = row.raw_description
        if not literal.strip():
            outer_census["blank"] += 1
            continue
        match = TSMIS_DESCRIPTION_RE.fullmatch(literal)
        if match is None:
            outer_census["no_numeric_slash_prefix"] += 1
            continue
        outer = ROUTE_RE.fullmatch(_text(match.group("route"), upper=True))
        current = ROUTE_RE.fullmatch(row.route)
        if outer is None or current is None or int(outer.group(1)) != int(current.group(1)):
            outer_census["different_numeric_prefix_preserved"] += 1
        elif outer.group(2) == current.group(2):
            outer_census["exact_outer_route_prefix_stripped"] += 1
        elif not outer.group(2) and current.group(2):
            outer_census["base_route_prefix_on_suffixed_route_stripped"] += 1
            if len(suffix_examples) < 20:
                suffix_examples.append({
                    "key": list(row.key), "reference": row.reference,
                    "route": row.route, "raw_description": row.raw_description,
                    "projected_description": row.description})
        else:
            outer_census["conflicting_suffix_prefix_preserved"] += 1
    return {
        "raw_tsn_numeric_prefix_count": len(prefixed),
        "all_15_excel_exact": (
            len(prefixed) == 15 and all(item["excel_exact"] for item in prefixed)),
        "all_15_pdf_render_exact": (
            len(prefixed) == 15 and all(item["pdf_render_exact"] for item in prefixed)),
        "prefixed_rows": prefixed,
        "tsmis_excel_outer_prefix_census": dict(sorted(outer_census.items())),
        "suffix_route_outer_prefix_examples": suffix_examples,
    }


def _tsmis_cross_format(excel_rows: Sequence[RampRow],
                        pdf_rows: Sequence[RampRow]) -> dict[str, object]:
    comparison = _compare(
        "TSMIS PDF vs TSMIS Excel", pdf_rows, excel_rows,
        BASE_ASSERTED_FIELDS, collapse_description=True,
        context_fields=("onoff", "ramp_type"))
    classifications: Counter[str] = Counter()
    examples: dict[str, list[dict[str, object]]] = defaultdict(list)
    excel_by_ref = {row.reference: row for row in excel_rows}
    pdf_by_ref = {row.reference: row for row in pdf_rows}
    # Pair references differ by member extension and page/row, so classify from the
    # already paired comparison details and independently indexed key groups.
    for item in comparison["differing_rows"]:
        for difference in item["differences"]:
            field = difference["field"]
            left = difference["left"]
            right = difference["right"]
            if field == "description" and "_x000d_" in str(right).lower():
                kind = "excel_literal_x000d_escape_absent_from_pdf"
            else:
                kind = "unclassified"
            classifications[kind] += 1
            if len(examples[kind]) < 20:
                examples[kind].append({
                    "key": item["key"], "field": field,
                    "pdf": left, "excel": right,
                    "pdf_reference": item["left"],
                    "excel_reference": item["right"],
                })
    raw_render = _classify_tsmis_render_equivalence(excel_rows, pdf_rows)
    return {
        "semantic_comparison": comparison,
        "semantic_difference_classification": {
            "counts": dict(sorted(classifications.items())),
            "examples": dict(examples),
            "all_classified": set(classifications) <= {
                "excel_literal_x000d_escape_absent_from_pdf"},
        },
        "raw_render_equivalence": raw_render,
    }


def _classify_tsmis_render_equivalence(excel_rows: Sequence[RampRow],
                                       pdf_rows: Sequence[RampRow]) -> dict[str, object]:
    pdf_by_key: dict[tuple[str, str, str], list[RampRow]] = defaultdict(list)
    excel_by_key: dict[tuple[str, str, str], list[RampRow]] = defaultdict(list)
    for row in pdf_rows:
        pdf_by_key[row.key].append(row)
    for row in excel_rows:
        excel_by_key[row.key].append(row)
    pairs: list[tuple[RampRow, RampRow]] = []
    for key in sorted(set(pdf_by_key) | set(excel_by_key)):
        selected, pdf_only, excel_only, _trace = _pair_group(
            pdf_by_key[key], excel_by_key[key], BASE_ASSERTED_FIELDS,
            collapse_description=True)
        if pdf_only or excel_only:
            raise AuditError(
                f"same-pull TSMIS render classifier found one-sided key {key!r}")
        pairs.extend(selected)
    counts: Counter[str] = Counter()
    samples: dict[str, list[dict[str, object]]] = defaultdict(list)
    whitespace_breakdown: Counter[str] = Counter()
    whitespace_equivalent_rows = 0

    def rendered_description(row: RampRow) -> str:
        raw = row.raw_description
        literal = raw.strip()
        if row.source == "TSMIS PDF" and literal == NULL_DESCRIPTION:
            return literal
        match = TSMIS_DESCRIPTION_RE.fullmatch(raw)
        if match is None:
            return raw
        outer = ROUTE_RE.fullmatch(_text(match.group("route"), upper=True))
        current = ROUTE_RE.fullmatch(row.route)
        if (outer is not None and current is not None
                and int(outer.group(1)) == int(current.group(1))
                and outer.group(2) in {"", current.group(2)}):
            return match.group("description")
        return raw

    for pdf, excel in pairs:
        if pdf.raw_area4 == "-" and excel.raw_area4.strip() == "":
            counts["pdf_dash_renders_excel_blank"] += 1
        elif pdf.area4 != excel.area4:
            counts["unclassified"] += 1
            samples["unclassified"].append({
                "key": list(pdf.key), "field": "area4",
                "pdf": pdf.raw_area4, "excel": excel.raw_area4,
                "pdf_reference": pdf.reference, "excel_reference": excel.reference})

        pdf_value = rendered_description(pdf)
        excel_value = rendered_description(excel)
        excel_without_escape = re.sub(
            r"_x000d_", "", excel_value, flags=re.IGNORECASE)
        has_escape = excel_without_escape != excel_value
        if (pdf.raw_description.strip() == NULL_DESCRIPTION
                and excel_value.strip() == ""):
            kind = "pdf_no_linear_event_renders_excel_blank"
        elif pdf_value == excel_value:
            continue
        elif has_escape and _collapse(pdf_value) == _collapse(excel_without_escape):
            kind = "excel_literal_x000d_escape_absent_from_pdf"
            whitespace_equivalent_rows += 1
            if pdf_value.strip() == excel_without_escape.strip():
                whitespace_breakdown["edge_only"] += 1
            else:
                whitespace_breakdown[
                    "internal_runs_or_internal_plus_edge"] += 1
        elif _collapse(pdf_value) == _collapse(excel_value):
            kind = "html_whitespace_collapse"
            whitespace_equivalent_rows += 1
            if pdf_value.strip() == excel_value.strip():
                whitespace_breakdown["edge_only"] += 1
            else:
                whitespace_breakdown[
                    "internal_runs_or_internal_plus_edge"] += 1
        else:
            kind = "unclassified"
        counts[kind] += 1
        if len(samples[kind]) < 20:
            samples[kind].append({
                "key": list(pdf.key), "field": "description",
                "pdf": pdf_value, "excel": excel_value,
                "pdf_reference": pdf.reference, "excel_reference": excel.reference})
    return {
        "counts": dict(sorted(counts.items())),
        "whitespace_equivalent_rows": whitespace_equivalent_rows,
        "whitespace_breakdown": dict(sorted(whitespace_breakdown.items())),
        "examples": dict(samples),
        "all_classified": "unclassified" not in counts,
        "paired_rows": len(pairs),
        "pairing_sha256": _canonical_digest(
            [(pdf.reference, excel.reference) for pdf, excel in pairs]),
    }


def _row_payload(row: RampRow) -> dict[str, object]:
    return {
        key: value for key, value in asdict(row).items()
        if key not in {"source", "member", "source_record"}
    }


def _row_payload_digest(rows: Sequence[RampRow]) -> str:
    return _canonical_digest([_row_payload(row) for row in rows])


def _inspect_product_consolidated(path: Path, *, pdf: bool,
                                  truth_rows: Sequence[RampRow]) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    parsed: list[RampRow] = []
    formulas = []
    errors = []
    expected_header = ("Route", *(TSMIS_PDF_HEADER if pdf else TSMIS_HEADER))
    try:
        if workbook.sheetnames != [TSMIS_SHEET]:
            raise AuditError(
                f"product consolidated sheet universe drift: {workbook.sheetnames!r}")
        worksheet = workbook[TSMIS_SHEET]
        physical = worksheet.iter_rows()
        header_cells = next(physical, ())
        header = tuple(cell.value for cell in header_cells)
        if header != expected_header:
            raise AuditError(f"product consolidated header drift: {header!r}")
        for source_row, cells in enumerate(physical, 2):
            values = tuple(cell.value for cell in cells)
            if len(values) > len(expected_header) or all(
                    value is None for value in values):
                raise AuditError(
                    f"product consolidated row {source_row} physical drift")
            # XLSX omits physically trailing blank cells.  The authentic shifted
            # TSMIS layout declares a final blank column, so unsized streaming
            # sheets may yield fewer cells until the next populated coordinate.
            # Pad only the trailing absence; source-payload equality below still
            # rejects every missing nonblank value and any extra physical column.
            values += (None,) * (len(expected_header) - len(values))
            for cell in cells:
                if cell.data_type == "f":
                    formulas.append(cell.coordinate)
                elif cell.data_type == "e":
                    errors.append((cell.coordinate, cell.value))
            route = _route(values[0])
            parsed.append(_row_from_tsmis_values(
                source="product consolidated", member=path.name,
                source_record=f"row {source_row}", file_route=route,
                values=values[1:], pdf=pdf))
    finally:
        workbook.close()
    if formulas or errors:
        raise AuditError(
            f"product consolidated formula/error cells: {formulas[:3]!r} "
            f"{errors[:3]!r}")
    observed_digest = _row_payload_digest(parsed)
    expected_digest = _row_payload_digest(truth_rows)
    if len(parsed) != len(truth_rows) or observed_digest != expected_digest:
        first_mismatch = next((
            index for index, (observed, expected) in enumerate(
                zip(parsed, truth_rows), 2)
            if _row_payload(observed) != _row_payload(expected)), None)
        raise AuditError(
            f"product consolidated projection drift: rows={len(parsed)}/"
            f"{len(truth_rows)} first={first_mismatch} digest="
            f"{observed_digest}/{expected_digest}")
    return {
        "rows": len(parsed),
        "columns": len(expected_header),
        "ordered_source_payload_sha256": observed_digest,
        "projection_exact": True,
        "formula_cells": 0,
        "error_cells": 0,
    }


def _zip_digest_without_core(path: Path) -> str:
    digest = hashlib.sha256()
    with zipfile.ZipFile(path) as archive:
        names = sorted(
            name for name in archive.namelist()
            if name != "docProps/core.xml")
        for name in names:
            payload = archive.read(name)
            digest.update(name.encode("utf-8"))
            digest.update(b"\0")
            digest.update(len(payload).to_bytes(8, "big"))
            digest.update(hashlib.sha256(payload).digest())
    return digest.hexdigest()


def _sheet_digest(worksheet) -> dict[str, object]:
    digest = hashlib.sha256()
    rows = 0
    header: tuple[object, ...] = ()
    for row in worksheet.iter_rows(values_only=True):
        values = tuple(row)
        if rows == 0:
            header = values
        digest.update(_canonical_bytes(values))
        digest.update(b"\n")
        rows += 1
    return {
        "physical_rows": rows,
        "header": list(header),
        "ordered_typed_sha256": digest.hexdigest(),
    }


def _helper_outcome_exact(label: str, payload: dict[str, object],
                          expected: dict[str, object]) -> dict[str, object]:
    result = payload.get("result")
    if not isinstance(result, dict):
        raise AuditError(f"product {label} outcome missing")
    counts = result.get("counts")
    if not isinstance(counts, dict):
        raise AuditError(f"product {label} counts missing")
    expected_counts = {
        "known": True,
        "paired_rows": expected["both_rows"],
        "side_a_only_rows": expected["side_a_only_rows"],
        "side_b_only_rows": expected["side_b_only_rows"],
        "differing_rows": expected["differing_rows"],
        "differing_cells": expected["differing_cells"],
    }
    if any(counts.get(key) != value for key, value in expected_counts.items()):
        raise AuditError(
            f"product {label} returned-count drift: {counts!r}")
    product_per_field = {
        str(key).split(":", 1)[-1]: value
        for key, value in (counts.get("per_field_counts") or {}).items()
    }
    if product_per_field != expected["per_field"]:
        raise AuditError(
            f"product {label} per-field outcome drift: {product_per_field!r}")
    if (result.get("status"), result.get("completion"),
            result.get("verdict"), result.get("skipped_inputs"),
            result.get("failed_inputs")) != ("ok", "complete", "diff", 0, 0):
        raise AuditError(f"product {label} outcome state drift")
    if result.get("warnings") or result.get("failures"):
        raise AuditError(f"product {label} reported warnings/failures")
    generation = result.get("artifact_generation")
    if not isinstance(generation, dict) or (
            generation.get("completion"), generation.get("publication_state"),
            generation.get("requested_mode")) != (
                "complete", "committed", "both"):
        raise AuditError(f"product {label} generation state drift")
    members = generation.get("members")
    flavors = sorted(
        str(member.get("flavor")) for member in members
        if isinstance(member, dict)) if isinstance(members, list) else []
    if flavors != ["formulas", "values"]:
        raise AuditError(f"product {label} twin manifest drift: {flavors!r}")
    return {
        "status": result.get("status"),
        "completion": result.get("completion"),
        "verdict": result.get("verdict"),
        "skipped_inputs": result.get("skipped_inputs"),
        "failed_inputs": result.get("failed_inputs"),
        "counts": counts,
        "artifact_generation": {
            "completion": generation.get("completion"),
            "publication_state": generation.get("publication_state"),
            "requested_mode": generation.get("requested_mode"),
            "flavors": flavors,
        },
    }


def _inspect_product_comparison(
        label: str, formulas_path: Path, values_path: Path,
        expected: dict[str, object],
        description_contract: dict[str, object]) -> dict[str, object]:
    side_a = str(expected["side_a"])
    side_b = str(expected["side_b"])
    expected_sheets = [
        "Summary", "Spot Check", "Comparison", "Routes",
        f"Only in {side_a}", f"Only in {side_b}", side_a, side_b,
        "Notes", "__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B",
    ]
    formulas = load_workbook(formulas_path, read_only=True, data_only=False)
    try:
        if formulas.sheetnames != expected_sheets:
            raise AuditError(
                f"product {label} formulas sheet universe drift: "
                f"{formulas.sheetnames!r}")
        formula_found = any(
            cell.data_type == "f"
            for row in formulas["Comparison"].iter_rows(min_row=2, max_row=25)
            for cell in row)
        if not formula_found:
            raise AuditError(f"product {label} formulas flavor has no formulas")
    finally:
        formulas.close()

    values = load_workbook(values_path, read_only=True, data_only=True)
    try:
        if values.sheetnames != expected_sheets:
            raise AuditError(
                f"product {label} values sheet universe drift: {values.sheetnames!r}")
        comparison = values["Comparison"]
        physical = comparison.iter_rows(values_only=True)
        header = tuple(next(physical, ()))
        fields = list(header[7:18])
        expected_fields = [
            "PR", "Date of Record", "HG", "Area 4", "City Code", "R/U",
            "Description", "Ramp Name", "On/Off", "Ramp Type", "ADT",
        ]
        expected_header = (
            "Route", "PM", "#", f"{side_a} Row", f"{side_b} Row",
            "Status", "Diffs", *expected_fields,
            "__CMP_E1_STATE_V1_C001_P0000_P0010",
        )
        if header != expected_header:
            raise AuditError(f"product {label} Comparison header drift: {header!r}")

        desired_prefix_keys = {
            (str(item["key"][0]), str(item["key"][2]))
            for item in description_contract["prefixed_rows"]
        } if "vs_tsn" in label else set()
        prefix_observations = []
        district_observations = []
        status_counts: Counter[str] = Counter()
        per_field: Counter[str] = Counter()
        differing_rows = 0
        differing_cells = 0
        union_rows = 0
        ordered = hashlib.sha256()
        for row in physical:
            if not row or all(value is None for value in row):
                continue
            if len(row) != len(expected_header):
                raise AuditError(f"product {label} Comparison row width drift")
            union_rows += 1
            ordered.update(("\t".join(
                "" if value is None else str(value) for value in row
            ) + "\n").encode("utf-8"))
            status = str(row[5])
            status_counts[status] += 1
            state = str(row[18])
            if len(state) != len(fields):
                raise AuditError(f"product {label} state-vector width drift")
            difference_count = state.count("D")
            if status == "Both":
                if row[6] != difference_count:
                    raise AuditError(
                        f"product {label} Diffs/state disagreement at {row[:2]!r}")
                differing_rows += int(difference_count > 0)
                differing_cells += difference_count
                for field, state_value in zip(fields, state):
                    per_field[field] += int(state_value == "D")
            weak_key = (str(row[0]), str(row[1]))
            if weak_key in desired_prefix_keys:
                description_index = fields.index("Description")
                prefix_observations.append({
                    "weak_key": list(weak_key),
                    "status": status,
                    "diffs": row[6],
                    "description": row[13],
                    "description_state": state[description_index],
                })
            if weak_key == ("005", "72.366") and "vs_tsn" in label:
                district_observations.append({
                    "weak_key": list(weak_key), "status": status,
                    "diffs": row[6], "state": state,
                    "description": row[13],
                })

        expected_statuses = {
            "Both": expected["both_rows"],
            f"{side_a} only": expected["side_a_only_rows"],
            f"{side_b} only": expected["side_b_only_rows"],
        }
        expected_statuses = {
            key: value for key, value in expected_statuses.items() if value or key == "Both"
        }
        if dict(status_counts) != expected_statuses:
            raise AuditError(
                f"product {label} status census drift: {dict(status_counts)!r}")
        if (union_rows != sum(expected_statuses.values())
                or differing_rows != expected["differing_rows"]
                or differing_cells != expected["differing_cells"]
                or {field: per_field[field] for field in fields}
                != expected["per_field"]
                or ordered.hexdigest() != expected["ordered_rows_sha256"]):
            raise AuditError(f"product {label} independently read truth drift")

        one_sided = {}
        for side, expected_count in (
                (side_a, expected["side_a_only_rows"]),
                (side_b, expected["side_b_only_rows"])):
            worksheet = values[f"Only in {side}"]
            rows = worksheet.iter_rows(values_only=True)
            side_header = tuple(next(rows, ()))
            count = sum(
                1 for row in rows if row and any(value is not None for value in row))
            if count != expected_count:
                raise AuditError(
                    f"product {label} Only-in-{side} count drift: {count}")
            one_sided[side] = {"rows": count, "header": list(side_header)}

        snapshot_a = _sheet_digest(values["__CMP_E2_SNAPSHOT_A"])
        snapshot_b = _sheet_digest(values["__CMP_E2_SNAPSHOT_B"])
        note_text = "\n".join(
            str(value)
            for row in values["Notes"].iter_rows(values_only=True)
            for value in row if value is not None)
        notes_declare_route_plus_pm = "Rows are keyed on Route + PM" in note_text
        if "vs_tsn" in label and not notes_declare_route_plus_pm:
            raise AuditError(f"product {label} Notes key statement drift")
    finally:
        values.close()

    if "vs_tsn" in label:
        if (len(prefix_observations) != 15
                or any(item["status"] != "Both"
                       or item["description_state"] != "D"
                       or " ≠ " not in str(item["description"])
                       for item in prefix_observations)):
            raise AuditError(
                f"product {label} exact 15 prefix-loss witness drift")
        if (len(district_observations) != 1
                or district_observations[0]["status"] != "Both"
                or district_observations[0]["diffs"] != 0
                or "D" in district_observations[0]["state"]):
            raise AuditError(
                f"product {label} District false-clean witness drift")

    return {
        "counts": {
            "union_rows": union_rows,
            "both_rows": status_counts["Both"],
            "side_a_only_rows": expected["side_a_only_rows"],
            "side_b_only_rows": expected["side_b_only_rows"],
            "differing_rows": differing_rows,
            "differing_cells": differing_cells,
            "per_field": {field: per_field[field] for field in fields},
        },
        "ordered_rows_sha256": ordered.hexdigest(),
        "comparison_identity_columns": ["Route", "PM"],
        "district_column_present": "District" in header,
        "county_column_present": "County" in header,
        "notes_declare_route_plus_pm_key": notes_declare_route_plus_pm,
        "one_sided_sheets": one_sided,
        "snapshots": {"side_a": snapshot_a, "side_b": snapshot_b},
        "exact_15_prefix_false_differences": prefix_observations,
        "district_12_vs_11_false_clean": district_observations,
        "formulas_package_without_core_sha256": _zip_digest_without_core(
            formulas_path),
        "values_package_without_core_sha256": _zip_digest_without_core(
            values_path),
    }


def _parse_helper_stdout(stdout: str) -> dict[str, object]:
    for line in reversed([line.strip() for line in stdout.splitlines() if line.strip()]):
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict) and "comparisons" in payload:
            return payload
    raise AuditError("Ramp Detail product witness emitted no structured result")


def _run_product(
        product_root: Path, xlsx_root: Path, pdf_root: Path,
        tsn_raw: Path, tsn_normalized: Path,
        excel_rows: Sequence[RampRow], pdf_rows: Sequence[RampRow],
        description_contract: dict[str, object]) -> dict[str, object]:
    completed = subprocess.run(
        [
            sys.executable, str(PRODUCT_HELPER_PATH),
            "--xlsx-root", str(xlsx_root), "--pdf-root", str(pdf_root),
            "--tsn-raw", str(tsn_raw),
            "--tsn-normalized", str(tsn_normalized),
            "--work-root", str(product_root),
        ],
        cwd=REPO_ROOT,
        env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        text=True, encoding="utf-8", errors="replace",
        capture_output=True, timeout=900, check=False)
    if completed.returncode != 0:
        raise AuditError(
            f"Ramp Detail product witness failed ({completed.returncode}): "
            f"{completed.stderr[-3000:]!r} {completed.stdout[-3000:]!r}")
    payload = _parse_helper_stdout(completed.stdout)
    consolidations = payload.get("consolidations")
    comparisons = payload.get("comparisons")
    if not isinstance(consolidations, dict) or not isinstance(comparisons, dict):
        raise AuditError("Ramp Detail product witness result shape drift")

    expected_consolidated = {
        "excel": product_root / "ramp_detail_excel_consolidated.xlsx",
        "pdf": product_root / "ramp_detail_pdf_consolidated.xlsx",
    }
    consolidated = {}
    for label, expected_path in expected_consolidated.items():
        item = consolidations.get(label)
        if not isinstance(item, dict) or (
                item.get("status"), item.get("completion"),
                item.get("skipped_inputs"), item.get("failed_inputs")) != (
                    "ok", "complete", 0, 0):
            raise AuditError(f"product {label} consolidation outcome drift")
        output = item.get("output")
        observed_path = Path(str(output.get("path"))).resolve() \
            if isinstance(output, dict) else Path()
        if observed_path != expected_path.resolve() or not observed_path.is_file():
            raise AuditError(f"product {label} consolidated path escaped/drifted")
        consolidated[label] = _inspect_product_consolidated(
            observed_path, pdf=(label == "pdf"),
            truth_rows=pdf_rows if label == "pdf" else excel_rows)

    inspected = {}
    helper_outcomes = {}
    for label, expected in EXPECTED_PRODUCT_COMPARISONS.items():
        item = comparisons.get(label)
        if not isinstance(item, dict):
            raise AuditError(f"product comparison {label} missing")
        outputs = item.get("outputs")
        if not isinstance(outputs, dict):
            raise AuditError(f"product comparison {label} output map missing")
        expected_paths = {
            "formulas": product_root / f"{label}.xlsx",
            "values": product_root / f"{label} (values).xlsx",
        }
        observed_paths = {}
        for flavor, expected_path in expected_paths.items():
            member = outputs.get(flavor)
            observed = Path(str(member.get("path"))).resolve() \
                if isinstance(member, dict) else Path()
            if observed != expected_path.resolve() or not observed.is_file():
                raise AuditError(
                    f"product {label} {flavor} path escaped/drifted")
            observed_paths[flavor] = observed
        helper_outcomes[label] = _helper_outcome_exact(label, item, expected)
        inspected[label] = _inspect_product_comparison(
            label, observed_paths["formulas"], observed_paths["values"],
            expected, description_contract)

    for prefix in ("excel", "pdf"):
        raw = inspected[f"{prefix}_vs_tsn_raw"]
        normalized = inspected[f"{prefix}_vs_tsn_normalized"]
        for key in ("counts", "ordered_rows_sha256", "comparison_identity_columns",
                    "district_column_present", "county_column_present",
                    "exact_15_prefix_false_differences",
                    "district_12_vs_11_false_clean", "snapshots"):
            if raw[key] != normalized[key]:
                raise AuditError(
                    f"product {prefix} raw/normalized semantic drift at {key}")

    product_code = payload.get("loaded_product_code")
    if not isinstance(product_code, dict) or not product_code.get("entries"):
        raise AuditError("Ramp Detail product loaded-code manifest missing")
    return {
        "consolidations": consolidated,
        "helper_outcomes": helper_outcomes,
        "comparisons": inspected,
        "raw_and_normalized_product_paths_semantically_identical": True,
        "loaded_product_code": product_code,
    }


def _loaded_oracle_module_manifest() -> dict[str, object]:
    entries = []
    for module_name, module in sorted(sys.modules.items()):
        if module_name.split(".", 1)[0] not in {
                "openpyxl", "pdfplumber", "pdfminer", "pypdf"}:
            continue
        raw = getattr(module, "__file__", None)
        if not raw:
            continue
        path = Path(raw).resolve()
        if path.suffix.lower() != ".py" or not path.is_file():
            continue
        entries.append({
            "module": module_name,
            "filename": path.name,
            "bytes": path.stat().st_size,
            "sha256": _sha_file(path),
        })
    return {
        "module_file_count": len(entries),
        "canonical_json_sha256": _canonical_digest(entries),
        "entries": entries,
    }


def _product_manifest_current(manifest: dict[str, object]) -> dict[str, object]:
    checks = []
    entries = manifest.get("entries")
    if not isinstance(entries, list):
        return {"all_current": False, "checks": [], "reason": "entries missing"}
    scripts_root = (REPO_ROOT / "scripts").resolve()
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        relative = str(entry.get("relative_path", ""))
        path = (scripts_root / relative).resolve()
        try:
            path.relative_to(scripts_root)
        except ValueError:
            observed = None
        else:
            observed = _file_identity(path) if path.is_file() else None
        expected = {"bytes": entry.get("bytes"), "sha256": entry.get("sha256")}
        checks.append({
            "module": entry.get("module"), "relative_path": relative,
            "expected": expected, "observed": observed,
            "current": observed == expected,
        })
    return {
        "all_current": bool(checks) and all(item["current"] for item in checks),
        "checks": checks,
    }


def _run_gate(path: Path) -> dict[str, object]:
    completed = subprocess.run(
        [sys.executable, str(path)], cwd=REPO_ROOT,
        env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        text=True, encoding="utf-8", errors="replace",
        capture_output=True, timeout=240, check=False)
    if completed.returncode != 0:
        raise AuditError(
            f"Ramp Detail Stage-8 mutation gate failed ({completed.returncode}): "
            f"{completed.stdout[-3000:]!r} {completed.stderr[-3000:]!r}")
    return {
        "status": "executed_pass",
        "gate_identity": _file_identity(path),
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
    }


def _source_digest_exact(label: str, summary: dict[str, object]) -> bool:
    expected = EXPECTED_SOURCE_DIGESTS[label]
    return (
        summary.get("rows") == expected["rows"]
        and summary.get("route_count") == expected["route_count"]
        and summary.get("ordered_row_sha256") == expected["ordered_row_sha256"]
        and summary.get("raw_representation_sha256")
        == expected["raw_representation_sha256"]
        and summary.get("identity", {}).get("unique_keys")
        == expected["unique_keys"]
        and summary.get("identity", {}).get("duplicate_groups")
        == expected["duplicate_groups"])


def _comparison_exact(summary: dict[str, object],
                      expected: dict[str, object]) -> bool:
    expected_counts = {
        key: value for key, value in expected.items()
        if key != "paired_ledger_sha256"
    }
    return (
        summary.get("counts") == expected_counts
        and summary.get("paired_ledger_sha256")
        == expected["paired_ledger_sha256"])


def _publication_current(args: argparse.Namespace,
                         result: dict[str, object]) -> tuple[
                             bool, dict[str, object]]:
    tree_paths = {
        "tsmis_xlsx": args.tsmis_xlsx_root,
        "tsmis_pdf": args.tsmis_pdf_root,
    }
    file_paths = {
        "tsn_xlsx": args.tsn_xlsx,
        "tsn_pdf": args.tsn_pdf,
        "tsn_normalized": args.tsn_normalized,
        "tsn_normalized_sidecar": args.tsn_normalized_sidecar,
        "stage6_result": args.stage6_result,
        "stage6_acceptance": args.stage6_acceptance,
        "tsn_cross_format": args.tsn_cross_format,
    }
    detail: dict[str, object] = {"trees": {}, "files": {}, "code": {}}
    captures = result["source_capture"]
    for label, path in tree_paths.items():
        observed, _entries = _manifest(path, str(TREE_BINDINGS[label]["suffix"]))
        expected = captures[label]["observed"]
        detail["trees"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    for label, path in file_paths.items():
        observed = _file_identity(path) if path.is_file() else None
        expected = captures[label]["observed"]
        detail["files"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    code_paths = {
        "generator": GENERATOR_PATH,
        "product_helper": PRODUCT_HELPER_PATH,
        "self_gate": SELF_GATE_PATH,
    }
    for label, path in code_paths.items():
        observed = _file_identity(path) if path.is_file() else None
        expected = result["provenance"]["code_identities"][label]
        detail["code"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    parser_manifest = _loaded_oracle_module_manifest()
    expected_parser = result["provenance"]["loaded_oracle_module_manifest"]
    detail["code"]["oracle_parser_modules"] = {
        "expected_sha256": expected_parser["canonical_json_sha256"],
        "observed_sha256": parser_manifest["canonical_json_sha256"],
        "expected_files": expected_parser["module_file_count"],
        "observed_files": parser_manifest["module_file_count"],
        "current": parser_manifest == expected_parser,
    }
    product_current = _product_manifest_current(
        result["production"]["loaded_product_code"])
    detail["code"]["loaded_product_modules"] = product_current
    flags = [
        item["current"]
        for group in (detail["trees"], detail["files"], detail["code"])
        for item in group.values()
        if isinstance(item, dict) and "current" in item
    ]
    flags.append(product_current["all_current"])
    return bool(flags) and all(flags), detail


def _cleanup_private_snapshot(snapshot_root: Path,
                              work_root: Path) -> dict[str, object]:
    snapshot_root = snapshot_root.resolve()
    work_root = work_root.resolve()
    if (snapshot_root.parent != work_root
            or not snapshot_root.name.startswith("ramp-detail-stage8-")):
        raise AuditError(
            f"refusing cleanup outside verified private root: {snapshot_root}")
    attempts = []
    for attempt in range(1, 21):
        gc.collect()
        try:
            shutil.rmtree(snapshot_root)
        except FileNotFoundError:
            return {"complete": True, "attempts": attempts, "residue": None}
        except (PermissionError, OSError) as exc:
            attempts.append({
                "attempt": attempt, "type": type(exc).__name__,
                "winerror": getattr(exc, "winerror", None),
            })
            time.sleep(0.5)
        else:
            return {"complete": True, "attempts": attempts, "residue": None}
    return {
        "complete": not snapshot_root.exists(),
        "attempts": attempts,
        "residue": str(snapshot_root) if snapshot_root.exists() else None,
    }


@contextmanager
def _private_snapshot_context(work_root: Path,
                              cleanup_state: dict[str, object]):
    snapshot_root = Path(tempfile.mkdtemp(
        prefix="ramp-detail-stage8-", dir=work_root)).resolve()
    if snapshot_root.parent != work_root.resolve():
        raise AuditError("private snapshot escaped the requested work root")
    try:
        yield snapshot_root
    finally:
        cleanup = _cleanup_private_snapshot(snapshot_root, work_root)
        cleanup_state.clear()
        cleanup_state.update(cleanup)
        if not cleanup["complete"] and sys.exc_info()[0] is None:
            raise AuditError(
                f"private snapshot cleanup failed: {cleanup['residue']}")


def _strip_rows(summary: dict[str, object]) -> list[RampRow]:
    return list(summary["rows_data"])  # type: ignore[arg-type]


def _public_source_summary(summary: dict[str, object]) -> dict[str, object]:
    return {key: value for key, value in summary.items() if key != "rows_data"}


def run(args: argparse.Namespace) -> dict[str, object]:
    code_identities = {
        "generator": _file_identity(GENERATOR_PATH),
        "product_helper": _file_identity(PRODUCT_HELPER_PATH),
        "self_gate": _file_identity(SELF_GATE_PATH),
    }
    mutation_gate = _run_gate(SELF_GATE_PATH)
    work_root = args.work_root.resolve()
    source_roots = [
        args.tsmis_xlsx_root.resolve(), args.tsmis_pdf_root.resolve(),
        args.tsn_xlsx.parent.resolve(), args.tsn_pdf.parent.resolve(),
        args.tsn_normalized.parent.resolve(),
    ]
    if any(work_root == source or work_root.is_relative_to(source)
           for source in source_roots):
        raise AuditError("private work root must not be inside a source tree")
    work_root.mkdir(parents=True, exist_ok=True)

    private_cleanup: dict[str, object] = {}
    with _private_snapshot_context(work_root, private_cleanup) as snapshot_root:
        source_capture: dict[str, dict[str, object]] = {}
        source_capture["tsmis_xlsx"], tsmis_xlsx_root = _capture_tree(
            "tsmis_xlsx", args.tsmis_xlsx_root, snapshot_root / "tsmis_xlsx")
        source_capture["tsmis_pdf"], tsmis_pdf_root = _capture_tree(
            "tsmis_pdf", args.tsmis_pdf_root, snapshot_root / "tsmis_pdf")
        source_capture["tsn_xlsx"], tsn_xlsx = _capture_file(
            "tsn_xlsx", args.tsn_xlsx,
            snapshot_root / "tsn_xlsx" / args.tsn_xlsx.name)
        source_capture["tsn_pdf"], tsn_pdf = _capture_file(
            "tsn_pdf", args.tsn_pdf,
            snapshot_root / "tsn_pdf" / args.tsn_pdf.name)
        source_capture["tsn_normalized"], tsn_normalized_path = _capture_file(
            "tsn_normalized", args.tsn_normalized,
            snapshot_root / "tsn_normalized" / args.tsn_normalized.name)
        source_capture[
            "tsn_normalized_sidecar"], normalized_sidecar_path = _capture_file(
                "tsn_normalized_sidecar", args.tsn_normalized_sidecar,
                snapshot_root / "tsn_normalized_sidecar"
                / args.tsn_normalized_sidecar.name)
        source_capture["stage6_result"], stage6_result = _capture_file(
            "stage6_result", args.stage6_result,
            snapshot_root / "stage6_result" / args.stage6_result.name)
        source_capture["stage6_acceptance"], stage6_acceptance = _capture_file(
            "stage6_acceptance", args.stage6_acceptance,
            snapshot_root / "stage6_acceptance" / args.stage6_acceptance.name)
        source_capture["tsn_cross_format"], tsn_cross_format_path = _capture_file(
            "tsn_cross_format", args.tsn_cross_format,
            snapshot_root / "tsn_cross_format" / args.tsn_cross_format.name)

        tsmis_excel = _parse_tsmis_xlsx(tsmis_xlsx_root)
        tsmis_pdf = _parse_tsmis_pdf(tsmis_pdf_root)
        tsn = _parse_tsn_xlsx(tsn_xlsx)
        tsn_normalized = _parse_tsn_normalized(tsn_normalized_path)
        normalized_sidecar = _load_normalized_sidecar(normalized_sidecar_path)
        stage6 = _load_stage6(stage6_result, stage6_acceptance)
        tsn_cross_format = _load_tsn_cross_format(tsn_cross_format_path)
        excel_rows = _strip_rows(tsmis_excel)
        pdf_rows = _strip_rows(tsmis_pdf)
        tsn_rows = _strip_rows(tsn)
        normalized_rows = _strip_rows(tsn_normalized)

        normalized_projection = _compare(
            "raw TSN XLSX vs normalized r7", tsn_rows, normalized_rows,
            ("district", "pr", "record_date", "hg", "area4", "city", "ru",
             "description", "ramp_name", "onoff", "ramp_type", "adt"),
            collapse_description=False)
        description_contract = _description_source_contract(
            excel_rows, pdf_rows, tsn_rows)
        cross_format = _tsmis_cross_format(excel_rows, pdf_rows)
        comparisons = {
            "tsmis_excel_vs_tsn": _compare(
                "TSMIS Excel vs TSN XLSX", excel_rows, tsn_rows,
                BASE_ASSERTED_FIELDS, collapse_description=False,
                context_fields=EXCEL_TSN_CONTEXT_FIELDS),
            "tsmis_pdf_vs_tsn": _compare(
                "TSMIS PDF vs TSN XLSX", pdf_rows, tsn_rows,
                PDF_TSN_ASSERTED_FIELDS, collapse_description=True,
                context_fields=PDF_TSN_CONTEXT_FIELDS),
        }
        weak_identity = {
            "tsmis_excel": _weak_identity_census(excel_rows),
            "tsmis_pdf": _weak_identity_census(pdf_rows),
            "tsn_excel": _weak_identity_census(tsn_rows),
            "tsn_normalized": _weak_identity_census(normalized_rows),
        }
        source_claims = _source_claim_contract(tsn_rows)
        production = _run_product(
            snapshot_root / "product", tsmis_xlsx_root, tsmis_pdf_root,
            tsn_xlsx, tsn_normalized_path, excel_rows, pdf_rows,
            description_contract)
        gc.collect()

    parser_manifest = _loaded_oracle_module_manifest()
    product_code_current = _product_manifest_current(
        production["loaded_product_code"])
    cross_semantic = cross_format["semantic_comparison"]
    excel_district = [
        row for row in comparisons["tsmis_excel_vs_tsn"]["differing_rows"]
        if any(diff["field"] == "district" for diff in row["differences"])]
    pdf_district = [
        row for row in comparisons["tsmis_pdf_vs_tsn"]["differing_rows"]
        if any(diff["field"] == "district" for diff in row["differences"])]
    exact_district_difference = {
        "key": ["005", "SD", "72.366"],
        "differences": [{"field": "district", "left": "12", "right": "11"}],
    }
    district_contract_exact = (
        len(excel_district) == len(pdf_district) == 1
        and excel_district[0]["key"] == exact_district_difference["key"]
        and pdf_district[0]["key"] == exact_district_difference["key"]
        and excel_district[0]["differences"]
        == exact_district_difference["differences"]
        and pdf_district[0]["differences"]
        == exact_district_difference["differences"])

    expected_left_only = [
        ["005", "LA", "25.218"], ["050", "SAC", "15.715"],
        ["050", "SAC", "15.823"], ["101", "LA", "1.284"],
    ]
    source_invariants = {
        "all_authoritative_source_members_match_bound_identities": all(
            capture["observed"] == capture["binding"]
            or all(capture["observed"].get(key) == capture["binding"].get(key)
                   for key in ("files", "bytes", "manifest_sha256"))
            for capture in source_capture.values()),
        "all_four_parsed_source_digests_exact": all((
            _source_digest_exact("tsmis_excel", tsmis_excel),
            _source_digest_exact("tsmis_pdf", tsmis_pdf),
            _source_digest_exact("tsn_excel", tsn),
            _source_digest_exact("tsn_normalized", tsn_normalized))),
        "all_source_route_universes_exact_126": (
            len(set(tsmis_excel["routes"])) == 126
            and set(tsmis_excel["routes"]) == set(tsmis_pdf["routes"])
            == set(tsn["routes"]) == set(tsn_normalized["routes"])),
        "tsmis_pdf_all_626_pages_500_data_pages_zero_residue": (
            tsmis_pdf["pages"] == 626 and tsmis_pdf["data_pages"] == 500
            and tsmis_pdf["unclassified_lines"] == 0
            and tsmis_pdf["unattached_description_fragments"] == 0
            and tsmis_pdf["unexpected_postmile_prefixes"] == 0),
        "normalized_outcome_sidecar_exact_complete": all(
            normalized_sidecar["checks"].values()),
        "accepted_stage6_raw_normalized_dependency_exact": all(
            stage6["checks"].values()),
        "accepted_tsn_xlsx_pdf_dependency_exact": all(
            tsn_cross_format["checks"].values()),
        "raw_normalized_exactly_15_description_losses": _comparison_exact(
            normalized_projection, EXPECTED_COMPARISONS["raw_vs_normalized"]),
        "all_15_numeric_descriptions_match_both_tsmis_formats": (
            description_contract["raw_tsn_numeric_prefix_count"] == 15
            and description_contract["all_15_excel_exact"]
            and description_contract["all_15_pdf_render_exact"]),
        "tsmis_outer_description_prefix_contract_exact": (
            description_contract["tsmis_excel_outer_prefix_census"] == {
                "base_route_prefix_on_suffixed_route_stripped": 46,
                "blank": 59,
                "exact_outer_route_prefix_stripped": 15_019,
                "no_numeric_slash_prefix": 92,
            }),
        "tsn_pm_suffix_source_claim_contract_exact": (
            source_claims["pm_suffix_nonblank"] == 313
            and source_claims["pm_suffix_counts"] == {"L": 165, "R": 148}
            and source_claims["all_nonblank_pm_suffix_equals_hg"]
            and not source_claims["blank_suffix_with_hg_l_or_r"]),
        "approved_d4_identity_and_weak_collision_census_exact": (
            weak_identity["tsn_excel"]["cross_county_weak_key_count"] == 81
            and weak_identity["tsn_excel"]["cross_county_identity_count"] == 163
            and _weak_identity_semantics(weak_identity["tsn_excel"])
            == _weak_identity_semantics(weak_identity["tsn_normalized"])
            and tsn["identity"]["unique_keys"] == 15_410),
        "tsmis_excel_vs_tsn_exact_truth": _comparison_exact(
            comparisons["tsmis_excel_vs_tsn"],
            EXPECTED_COMPARISONS["tsmis_excel_vs_tsn"]),
        "tsmis_pdf_vs_tsn_exact_truth": _comparison_exact(
            comparisons["tsmis_pdf_vs_tsn"],
            EXPECTED_COMPARISONS["tsmis_pdf_vs_tsn"]),
        "tsmis_pdf_vs_excel_exact_truth": _comparison_exact(
            cross_semantic, EXPECTED_COMPARISONS["tsmis_pdf_vs_excel"]),
        "exact_four_tsmis_only_physical_identities": (
            [row["key"] for row in comparisons[
                "tsmis_excel_vs_tsn"]["left_only"]] == expected_left_only
            and [row["key"] for row in comparisons[
                "tsmis_pdf_vs_tsn"]["left_only"]] == expected_left_only),
        "exact_one_real_district_difference": district_contract_exact,
        "exact_duplicate_assignment_contract": (
            comparisons["tsmis_excel_vs_tsn"]["duplicate_pairing"][
                "group_count"] == 1
            and comparisons["tsmis_pdf_vs_tsn"]["duplicate_pairing"][
                "group_count"] == 1
            and cross_semantic["duplicate_pairing"]["group_count"] == 1),
        "tsmis_cross_format_all_render_classes_exact": (
            cross_format["raw_render_equivalence"]["counts"] == {
                "excel_literal_x000d_escape_absent_from_pdf": 4,
                "html_whitespace_collapse": 306,
                "pdf_dash_renders_excel_blank": 59,
                "pdf_no_linear_event_renders_excel_blank": 59,
            }
            and cross_format["raw_render_equivalence"][
                "whitespace_equivalent_rows"] == 310
            and cross_format["raw_render_equivalence"][
                "whitespace_breakdown"] == {
                    "edge_only": 31,
                    "internal_runs_or_internal_plus_edge": 279,
                }
            and cross_format["raw_render_equivalence"]["all_classified"]
            and cross_format["semantic_difference_classification"][
                "counts"] == {"excel_literal_x000d_escape_absent_from_pdf": 4}
            and cross_format["semantic_difference_classification"][
                "all_classified"]),
    }
    production_invariants = {
        "permanent_mutation_gate_executed_pass": (
            mutation_gate["status"] == "executed_pass"),
        "production_tsmis_excel_pdf_consolidations_source_exact": all(
            item["projection_exact"]
            for item in production["consolidations"].values()),
        "production_all_five_formula_value_legs_independently_read_exact": (
            set(production["comparisons"]) == set(EXPECTED_PRODUCT_COMPARISONS)),
        "production_raw_and_normalized_paths_semantically_identical": production[
            "raw_and_normalized_product_paths_semantically_identical"],
        "production_exact_15_false_description_differences_reproduced": all(
            len(production["comparisons"][label][
                "exact_15_prefix_false_differences"]) == 15
            for label in (
                "excel_vs_tsn_raw", "excel_vs_tsn_normalized",
                "pdf_vs_tsn_raw", "pdf_vs_tsn_normalized")),
        "production_exact_district_false_clean_reproduced": all(
            len(production["comparisons"][label][
                "district_12_vs_11_false_clean"]) == 1
            for label in (
                "excel_vs_tsn_raw", "excel_vs_tsn_normalized",
                "pdf_vs_tsn_raw", "pdf_vs_tsn_normalized")),
        "production_weak_route_pm_identity_explicit": all(
            production["comparisons"][label][
                "comparison_identity_columns"] == ["Route", "PM"]
            and not production["comparisons"][label]["county_column_present"]
            for label in production["comparisons"]),
        "loaded_product_code_hash_manifest_current_at_result_build": (
            product_code_current["all_current"]),
        "loaded_oracle_parser_manifest_nonempty": (
            parser_manifest["module_file_count"] > 0),
        "private_snapshot_cleanup_complete": private_cleanup["complete"],
    }
    audit_invariants = {**source_invariants, **production_invariants}
    source_truth_exact = all(source_invariants.values())
    findings = {
        "oracle_blocking": [],
        "product_red": [
            {
                "finding": "CMP-AUD-045",
                "fact": (
                    "All Ramp Detail product triangle legs key on Route+PM and "
                    "discard County although the approved identity is "
                    "(Route, County, norm_pm(PM)); the exact TSN source has 81 "
                    "weak keys spanning 163 county identities."),
                "evidence": weak_identity["tsn_excel"],
            },
            {
                "finding": "CMP-AUD-135",
                "fact": (
                    "Raw and normalized production paths both create exactly 15 "
                    "false Description differences by deleting authoritative "
                    "leading numeric source data."),
                "evidence": production["comparisons"][
                    "excel_vs_tsn_normalized"][
                        "exact_15_prefix_false_differences"],
            },
            {
                "finding": "CMP-AUD-185",
                "fact": (
                    "The product omits District from every Ramp Detail comparison "
                    "and reports the exact 005/SD/72.366 District 12-vs-11 source "
                    "difference as a fully identical row."),
                "evidence": {
                    "source_truth": exact_district_difference,
                    "product": production["comparisons"][
                        "excel_vs_tsn_normalized"][
                            "district_12_vs_11_false_clean"],
                },
            },
            {
                "finding": "CMP-AUD-133",
                "fact": (
                    "The accepted normalized/comparison shape still omits the raw "
                    "PM_SFX, ADT_EFF_YEAR, and EFF_DATE claims needed for complete "
                    "source and evidence reconstruction."),
                "evidence": {
                    "source_claims": source_claims,
                    "stage6": stage6,
                },
            },
        ],
    }
    result = {
        "schema_version": 1,
        "audit": (
            "Stage 8 Ramp Detail authoritative four-source comparison oracle"),
        "methodology": {
            "authority": (
                "Exact same-pull 126-route TSMIS Excel and PDF trees, exact raw "
                "TSN XLSX and 500-page TSN print, accepted r7 normalized chain, "
                "and accepted TSN cross-format oracle."),
            "independence": (
                "Truth parsing imports no application parser, normalizer, schema, "
                "consolidator, comparator, or writer. Product imports occur only "
                "inside an isolated child after truth is derived."),
            "identity": ["Route", "County", "norm_pm(PM)"],
            "asserted_excel_vs_tsn": list(BASE_ASSERTED_FIELDS),
            "asserted_pdf_vs_tsn": list(PDF_TSN_ASSERTED_FIELDS),
            "outcomes_separated": [
                "source_truth_exact", "production_tsmis_projection_exact",
                "production_value_projection_exact",
                "production_comparison_semantics_exact",
                "stage8_base_oracle_complete",
                "comparison_end_to_end_perfect",
            ],
        },
        "bindings": {
            "source_trees": TREE_BINDINGS,
            "files": FILE_BINDINGS,
            "expected_source_digests": EXPECTED_SOURCE_DIGESTS,
            "expected_comparisons": EXPECTED_COMPARISONS,
            "expected_product_comparisons": EXPECTED_PRODUCT_COMPARISONS,
        },
        "source_capture": source_capture,
        "sources": {
            "tsmis_excel": _public_source_summary(tsmis_excel),
            "tsmis_pdf": _public_source_summary(tsmis_pdf),
            "tsn_excel": _public_source_summary(tsn),
            "tsn_normalized": _public_source_summary(tsn_normalized),
        },
        "accepted_dependencies": {
            "normalized_outcome": normalized_sidecar,
            "stage6_raw_to_normalized": stage6,
            "tsn_xlsx_to_pdf": tsn_cross_format,
        },
        "physical_identity": weak_identity,
        "source_claims": source_claims,
        "description_source_contract": description_contract,
        "raw_to_normalized_projection": normalized_projection,
        "tsmis_cross_format": cross_format,
        "comparisons": comparisons,
        "production": production,
        "dependency_gates": {
            "stage8_ramp_detail_mutations": mutation_gate,
        },
        "private_snapshot_cleanup": private_cleanup,
        "provenance": {
            "code_identities": code_identities,
            "loaded_oracle_module_manifest": parser_manifest,
            "loaded_product_code_current_at_result_build": product_code_current,
        },
        "findings": findings,
        "audit_invariants": audit_invariants,
        "source_truth_exact": source_truth_exact,
        "production_tsmis_projection_exact": all(
            item["projection_exact"]
            for item in production["consolidations"].values()),
        "production_value_projection_exact": False,
        "production_comparison_semantics_exact": False,
        "stage8_base_oracle_complete": all(audit_invariants.values()),
        "comparison_end_to_end_perfect": False,
    }
    current, current_detail = _publication_current(args, result)
    result["provenance"]["final_revalidation_at_result_build"] = {
        "all_current": current, "detail": current_detail,
    }
    result["audit_invariants"][
        "sources_and_code_current_at_result_build"] = current
    result["source_truth_exact"] = bool(result["source_truth_exact"] and current)
    result["stage8_base_oracle_complete"] = all(
        result["audit_invariants"].values())
    return result


def _atomic_write_text(path: Path, payload: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
                mode="w", encoding="utf-8", newline="\n", delete=False,
                dir=path.parent, prefix=f".{path.name}.", suffix=".tmp") as handle:
            temporary = Path(handle.name)
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary is not None:
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass


def _unlink_if_present(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass


def _write_decision(path: Path, output: Path, result: dict[str, object],
                    *, accepted: bool, reason: str,
                    postwrite_current: bool,
                    postwrite_detail: dict[str, object],
                    open_findings_authorized: bool) -> dict[str, object]:
    identity = _file_identity(output)
    decision = {
        "schema_version": 1,
        "accepted": accepted,
        "reason": reason,
        "audit": result.get("audit"),
        "result": str(output.resolve()),
        "result_bytes": identity["bytes"],
        "result_sha256": identity["sha256"],
        "source_truth_exact": result.get("source_truth_exact", False),
        "production_tsmis_projection_exact": result.get(
            "production_tsmis_projection_exact", False),
        "production_value_projection_exact": result.get(
            "production_value_projection_exact", False),
        "production_comparison_semantics_exact": result.get(
            "production_comparison_semantics_exact", False),
        "stage8_base_oracle_complete": result.get(
            "stage8_base_oracle_complete", False),
        "comparison_end_to_end_perfect": result.get(
            "comparison_end_to_end_perfect", False),
        "open_product_findings_authorized": open_findings_authorized,
        "post_result_write_revalidation": postwrite_current,
        "post_result_write_identities": postwrite_detail,
    }
    _atomic_write_text(path, json.dumps(
        decision, indent=2, ensure_ascii=False) + "\n")
    return decision


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--tsmis-xlsx-root", type=Path, default=DEFAULT_TSMIS_XLSX_ROOT)
    parser.add_argument("--tsmis-pdf-root", type=Path, default=DEFAULT_TSMIS_PDF_ROOT)
    parser.add_argument("--tsn-xlsx", type=Path, default=DEFAULT_TSN_XLSX)
    parser.add_argument("--tsn-pdf", type=Path, default=DEFAULT_TSN_PDF)
    parser.add_argument("--tsn-normalized", type=Path, default=DEFAULT_TSN_NORMALIZED)
    parser.add_argument(
        "--tsn-normalized-sidecar", type=Path,
        default=DEFAULT_TSN_NORMALIZED_SIDECAR)
    parser.add_argument("--stage6-result", type=Path, default=DEFAULT_STAGE6_RESULT)
    parser.add_argument(
        "--stage6-acceptance", type=Path, default=DEFAULT_STAGE6_ACCEPTANCE)
    parser.add_argument(
        "--tsn-cross-format", type=Path, default=DEFAULT_TSN_CROSS_FORMAT)
    parser.add_argument("--work-root", type=Path, default=DEFAULT_WORK_ROOT)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--allow-open-findings", action="store_true",
        help=(
            "accept the completed source/product witness while the exact "
            "documented product findings remain open"))
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    acceptance_path = args.output.with_suffix(
        args.output.suffix + ".acceptance.json")
    rejection_path = args.output.with_suffix(
        args.output.suffix + ".rejection.json")
    _unlink_if_present(acceptance_path)
    _unlink_if_present(rejection_path)
    try:
        result = run(args)
    except Exception as exc:
        failure = {
            "schema_version": 1,
            "audit": (
                "Stage 8 Ramp Detail authoritative four-source comparison oracle"),
            "error": {"type": type(exc).__name__, "message": str(exc)},
            "source_truth_exact": False,
            "production_tsmis_projection_exact": False,
            "production_value_projection_exact": False,
            "production_comparison_semantics_exact": False,
            "stage8_base_oracle_complete": False,
            "comparison_end_to_end_perfect": False,
        }
        _atomic_write_text(args.output, json.dumps(
            failure, indent=2, ensure_ascii=False) + "\n")
        decision = _write_decision(
            rejection_path, args.output, failure, accepted=False,
            reason="oracle_execution_failed", postwrite_current=False,
            postwrite_detail={}, open_findings_authorized=False)
        print(json.dumps({
            "accepted": False, "reason": decision["reason"],
            "output": str(args.output), "rejection": str(rejection_path),
            "error": failure["error"],
        }, ensure_ascii=False))
        return 2

    prewrite_current, prewrite_detail = _publication_current(args, result)
    result["publication_revalidation"] = {
        "after_complete_result_build": True,
        "before_result_write_all_current": prewrite_current,
        "before_result_write_identities": prewrite_detail,
    }
    result["audit_invariants"][
        "publication_inputs_current_before_write"] = prewrite_current
    result["stage8_base_oracle_complete"] = all(
        result["audit_invariants"].values())
    _atomic_write_text(args.output, json.dumps(
        result, indent=2, ensure_ascii=False) + "\n")

    postwrite_current, postwrite_detail = _publication_current(args, result)
    if not postwrite_current:
        result["publication_revalidation"][
            "post_result_write_all_current"] = False
        result["publication_revalidation"][
            "post_result_write_identities"] = postwrite_detail
        result["stage8_base_oracle_complete"] = False
        _atomic_write_text(args.output, json.dumps(
            result, indent=2, ensure_ascii=False) + "\n")
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False,
            reason="post_result_write_revalidation_failed",
            postwrite_current=False, postwrite_detail=postwrite_detail,
            open_findings_authorized=False)
        print(json.dumps({
            "accepted": False, "reason": decision["reason"],
            "rejection": str(rejection_path),
        }, ensure_ascii=False))
        return 2

    open_findings = bool(result["findings"]["product_red"])
    accepted = bool(
        result["source_truth_exact"]
        and result["production_tsmis_projection_exact"]
        and result["stage8_base_oracle_complete"]
        and postwrite_current
        and (not open_findings or args.allow_open_findings))
    if not accepted:
        reason = (
            "open_product_findings_not_authorized"
            if (result["stage8_base_oracle_complete"] and open_findings
                and not args.allow_open_findings)
            else "audit_or_projection_incomplete")
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False,
            reason=reason, postwrite_current=postwrite_current,
            postwrite_detail=postwrite_detail,
            open_findings_authorized=False)
        print(json.dumps({
            "accepted": False, "reason": reason,
            "output": str(args.output),
            "result_bytes": decision["result_bytes"],
            "result_sha256": decision["result_sha256"],
            "rejection": str(rejection_path),
        }, ensure_ascii=False))
        return 1 if reason == "open_product_findings_not_authorized" else 2

    decision = _write_decision(
        acceptance_path, args.output, result, accepted=True,
        reason="oracle_complete_with_documented_product_findings",
        postwrite_current=postwrite_current,
        postwrite_detail=postwrite_detail,
        open_findings_authorized=bool(args.allow_open_findings and open_findings))
    acceptance_identity = _file_identity(acceptance_path)
    print(json.dumps({
        "accepted": True,
        "output": str(args.output),
        "result_bytes": decision["result_bytes"],
        "result_sha256": decision["result_sha256"],
        "acceptance": str(acceptance_path),
        "acceptance_bytes": acceptance_identity["bytes"],
        "acceptance_sha256": acceptance_identity["sha256"],
        "source_truth_exact": result["source_truth_exact"],
        "production_tsmis_projection_exact": result[
            "production_tsmis_projection_exact"],
        "production_value_projection_exact": result[
            "production_value_projection_exact"],
        "production_comparison_semantics_exact": result[
            "production_comparison_semantics_exact"],
        "stage8_base_oracle_complete": result["stage8_base_oracle_complete"],
        "product_findings": len(result["findings"]["product_red"]),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
