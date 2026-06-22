"""Golden check for the cross-environment Highway Log (PDF) adapter
(compare_env.HIGHWAY_LOG_PDF) — v0.17.0. Both env sides are parsed from the app's
own PDF export (the accurate Highway Log source), via a flat_pdf_loader.

Locks: the adapter wiring (flat_pdf_loader set; sheet_name + force_header = the
corrected Highway Log header; base = the cross-env HL schema), its registration as a
folders/env row that promotes the HL-PDF matrix row's env mode from greyed to
supported (in COMPARE_REPORTS + matrix_rows + matrix._row_modes); and end-to-end that
the flat-PDF path yields a Route+Location-keyed Highway Log comparison (has_route=True)
that flags a genuine cell diff. The real PDF parsing is golden-tested in
check_tsmis_pdf_parse / check_tsmis_pdf_reconcile; here a stub flat_pdf_loader
exercises the EnvCompare PDF-flat machinery (CI-safe — no PDFs).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_env_highway_log_pdf.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
import highway_log_columns as hlc
import matrix
import reports
from events import Events
from openpyxl import load_workbook

_fail = []
DIFF = " ≠ "


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def test_wiring():
    print("wiring + matrix env un-grey:")
    a = compare_env.HIGHWAY_LOG_PDF
    check("flat_pdf_loader set (PDF-sourced flat path)",
          a.flat_pdf_loader is compare_env._load_highway_log_pdf_side
          and a.side_loader is None)
    check("sheet_name + force_header = corrected Highway Log header",
          a.sheet_name == compare_env._hl.SHEET_NAME
          and a.force_header == hlc.HEADER)
    check("registered as a folders/env row",
          any(adapter is a and kind == "folders" and group == "env"
              for _l, adapter, kind, group in reports.COMPARE_REPORTS))
    check("highway_log_pdf is a matrix row",
          "highway_log_pdf" in [r[0] for r in reports.matrix_rows()])
    defs = matrix._row_defs()
    hp = {m["id"]: m for m in matrix._row_modes("highway_log_pdf", "highway_log_pdf",
                                                defs["highway_log_pdf"][3])}
    check("matrix HL-PDF env mode now SUPPORTED (was greyed)", hp["env"]["supported"])
    check("HL-PDF row has_route True (flat HL shape)", defs["highway_log_pdf"][4] is True)


def test_flat_pdf_compare():
    print("flat-PDF env compare (Route+Location keyed; a genuine cell diff):")
    H = hlc.HEADER
    MARK = 6                                     # a plain (non-medwid/date/ditto) field

    def row(loc, marker):
        vals = [loc] + ["0"] * (len(H) - 1)     # v0 = Location (the key)
        vals[MARK] = marker
        return ["001"] + vals                   # [route, *H]

    def stub(folder, label, events):            # noqa: ARG001
        is_b = "ars" in str(folder).lower()
        return ([row("000.100", "1"), row("000.200", "9" if is_b else "1")], H, [])

    adapter = compare_env.EnvCompare(
        "highway_log_pdf", "Highway Log (PDF)", "highway_log_pdf",
        sheet_name=compare_env._hl.SHEET_NAME, base_schema=compare_env._HL_BASE,
        force_header=H, flat_pdf_loader=stub)
    root = Path(tempfile.mkdtemp(prefix="hlpdfenv_"))
    da, db = root / "2026-06-19 ssor-prod", root / "2026-06-19 ars-prod"
    da.mkdir(parents=True); db.mkdir(parents=True)
    out = root / "cmp.xlsx"
    res = adapter.compare_folders(str(da), str(db), str(out), events=Events(),
                                  confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    wb = load_workbook(out, read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        it = wb["Comparison"].iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
    finally:
        wb.close()
    check("env side sheets present (SSOR-PROD / ARS-PROD)",
          "SSOR-PROD" in sheets and "ARS-PROD" in sheets and "Comparison" in sheets)
    check("Location is the key column", "Location" in header)
    loc = header.index("Location")
    by = {r[loc]: r for r in rows}
    check("both locations matched (Route+Location keyed)",
          "000.100" in by and "000.200" in by)
    mark_col = header.index(H[MARK])
    check("perturbed field at 000.200 is a genuine diff", DIFF in by["000.200"][mark_col])
    check("unchanged 000.100 has no diff", DIFF not in " ".join(by["000.100"]))


def main():
    test_wiring()
    test_flat_pdf_compare()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-ENV-HIGHWAY-LOG-PDF CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
