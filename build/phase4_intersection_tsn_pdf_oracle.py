"""Independent Phase-4 Intersection TSN Excel/PDF parity oracle.

This deliberately does not import any production comparator, normalizer, evidence
adapter, schema, or report-view constant.  The source workbook and the two TSN
prints are treated as the only authorities.  The oracle proves:

* every two-line detail-print record maps to one physical workbook identity;
* all 36 workbook fields map to a fixed printed cell and are compared;
* complete postmile prefix identity prevents the six known prefix-only collisions;
* exact duplicates are conserved as multisets instead of being arbitrarily paired;
* the printed Summary categories are independently recomputed from detail rows; and
* the two physical lines expose an explicit source-to-Report-View mapping.

Cross-format differences are classified as exact, render-equivalent, an explicit
source-export delta, or unresolved.  Any unclassified row/cell loss makes the
process fail after writing its JSON result.
"""

from __future__ import annotations

import argparse
import collections
import dataclasses
import datetime as dt
import hashlib
import itertools
import json
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


DEFAULT_ROOT = Path(r"C:\Users\Yunus\Downloads\TSMIS\tsn_library")
DEFAULT_XLSX = DEFAULT_ROOT / "intersection_detail" / "raw" / "TSAR - INTERSECTION DETAIL_TSN.xlsx"
DEFAULT_DETAIL_PDF = DEFAULT_ROOT / "intersection_detail" / "pdf" / "Intersection Detail Statewide_TSN.pdf"
DEFAULT_SUMMARY_PDF = DEFAULT_ROOT / "intersection_summary" / "raw" / "Intersection Summary Statewide_TSN.pdf"

# Immutable acceptance sources.  A different byte stream is a different audit and
# must receive its own reviewed contract instead of inheriting this result.
EXPECTED_SOURCES = {
    "detail_xlsx": {
        "size": 2_920_705,
        "sha256": "5170ab19b957ba78ab0f175571f3aab51e8c49cac13fa307b3d0beaa023c84a2",
    },
    "detail_pdf": {
        "size": 9_284_543,
        "sha256": "1230b955176a1a34223ce8f79eeeed1b46970031372acc510ffb78a45c2f1f46",
    },
    "summary_pdf": {
        "size": 12_326,
        "sha256": "c3ad85848764df1b6da53c0bba0f785b3c045e83675f5983555ef514688a7d46",
    },
}
EXPECTED_TIMING = {
    "detail_pdf_creation": "D:20250915164913",
    "detail_xlsx_created": "2025-09-15 19:38:52",
}
BOUND_DETAIL_DELTA = {
    "identity": ("001", "VEN", "", "23.907"),
    "field": "DESCRIPTION",
    "excel": "SAN MIGUELITO RD-RT",
    "pdf": "A LEASE CANYON RD",
}

XLSX_HEADER = (
    "PP", "POST_MILE", "LOCATION", "DATE_REC", "HG", "CITY_CODE", "RU",
    "EFF_DATE_INT", "TY_INT", "EFF_DATE_CT", "TY_CT", "EFF_DATE_LT", "LT_TY",
    "EFF_DATE_ML", "MAIN_SM", "MAIN_LC", "MAIN_RC", "MAIN_TF", "MAIN_NL",
    "X_CROSS_OVERRIDE", "MAIN_EFF_DATE", "MAIN_ADT", "DESCRIPTION", "MAIN_OVERRIDE",
    "CROSS_BEGIN_DATE", "CS_SM", "CS_LC", "CS_RC", "CS_TF", "CS_NL", "EFF_DATE",
    "CROSS_ADT", "CROSS_ROUTE_NAME", "CROSS_PM_PREFIX", "CROSS_POSTMILE",
    "CROSS_PM_SUFFIX",
)

# Independently measured fixed-width cells in the 2025-09-15 Oracle print.  Bounds
# are PDF points and are validated by full-record conservation plus every-field
# parity, not accepted merely because a sample happens to parse.
LINE1_WINDOWS = (
    ("PP", 10, 24), ("POST_MILE", 24, 62), ("LOCATION", 64, 148),
    ("DATE_REC", 150, 200), ("HG", 202, 213), ("CITY_CODE", 213, 246),
    ("RU", 256, 278), ("EFF_DATE_INT", 279, 331), ("TY_INT", 332, 348),
    ("EFF_DATE_CT", 349, 394), ("TY_CT", 395, 412),
    ("EFF_DATE_LT", 413, 460), ("LT_TY", 473, 491),
    ("EFF_DATE_ML", 492, 541), ("MAIN_SM", 542, 555),
    ("MAIN_LC", 556, 567), ("MAIN_RC", 568, 578), ("MAIN_TF", 579, 588),
    ("MAIN_NL", 588, 599), ("X_CROSS_OVERRIDE", 599, 623),
    ("MAIN_EFF_DATE", 624, 673), ("MAIN_ADT", 674, 716),
)
LINE2_WINDOWS = (
    ("DESCRIPTION", 64, 251), ("MAIN_OVERRIDE", 251, 275),
    ("CROSS_BEGIN_DATE", 279, 331), ("CS_SM", 331, 345),
    ("CS_LC", 345, 357), ("CS_RC", 357, 369), ("CS_TF", 369, 381),
    ("CS_NL", 381, 395), ("EFF_DATE", 410, 458),
    ("CROSS_ADT", 458, 493), ("CROSS_ROUTE_NAME", 493, 514),
    ("PRINT_CROSS_ROUTE_SUFFIX", 514, 532),
    ("CROSS_PM_PREFIX", 532, 542), ("CROSS_POSTMILE", 542, 579),
    ("CROSS_PM_SUFFIX", 579, 598), ("PRINT_XING_ROUTE_SUFFIX", 598, 640),
)

LINE1_FIELDS = tuple(n for n, _a, _b in LINE1_WINDOWS)
LINE2_FIELDS = tuple(n for n, _a, _b in LINE2_WINDOWS if not n.startswith("PRINT_"))
assert set(LINE1_FIELDS + LINE2_FIELDS) == set(XLSX_HEADER)

DATE_FIELDS = {
    "DATE_REC", "EFF_DATE_INT", "EFF_DATE_CT", "EFF_DATE_LT", "EFF_DATE_ML",
    "MAIN_EFF_DATE", "CROSS_BEGIN_DATE", "EFF_DATE",
}
NUMERIC_FIELDS = {
    "POST_MILE", "MAIN_NL", "X_CROSS_OVERRIDE", "MAIN_ADT", "MAIN_OVERRIDE",
    "CS_NL", "CROSS_ADT", "CROSS_ROUTE_NAME", "CROSS_POSTMILE",
}
_LOC_RE = re.compile(r"^(\d{2})\s+([A-Z]{1,4})\.?\s+(\d{1,3})([A-Z]?)$")
_PM_RE = re.compile(r"^\d{3}\.\d{3}$")
_PDF_FLAGGED_DATE_RE = re.compile(r"^[*Y](\d{2}-\d{2}-\d{2})$")

# The bound Oracle Reports DESCRIPTION cell is the fixed 64..251 point cell in
# LINE2_WINDOWS.  Direct extraction of the hash-bound print establishes a
# 32-character capacity.  A source value is a render-only truncation only when
# the PDF reaches that exact measured boundary; a merely long shared prefix is
# not enough.
DESCRIPTION_PRINT_CONTRACT = {
    "cell_bounds_points": [64, 251],
    "font_family_suffix": "Courier",
    "characters": 32,
}

EXPECTED_PREFIX_PAIRS = {
    ("101", "SF", "5.45"): {"", "M"},
    ("115", "IMP", "9.54"): {"", "L"},
    ("132", "STA", "15.34"): {"", "L"},
    ("132", "STA", "15.62"): {"", "L"},
    ("184", "KER", "0"): {"", "L"},
    ("218", "MON", "0.34"): {"", "L"},
}

SUMMARY_CODES = {
    "HIGHWAY GROUP": ("R", "L", "X", "U", "D"),
    "RURAL/URBAN/SUBURBAN": ("R", "R-O", "U", "U-O", "+"),
    "INTERSECTION TYPE": ("F", "S", "Y", "M", "T", "Z"),
    "LIGHTING TYPE": ("N", "Y", "+"),
    "CONTROL TYPES": ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "P", "Z", "+"),
    "MAINLINE NUM OF LANES": ("1", "2", "3", "4", "5", "6", "7", "8", "+"),
    "MAINLINE MASTARM": ("Y", "N", "+"),
    "MAINLINE LEFT CHANNELIZATION": ("C", "N", "P", "R", "+"),
    "MAINLINE RIGHT CHANNELIZATION": ("Y", "N", "+"),
    "MAINLINE TRAFFIC FLOW": ("N", "P", "R", "W", "Z", "+"),
}
SIGNAL_CONTROL_CODES = frozenset("JKLMNP")


@dataclasses.dataclass(frozen=True)
class Identity:
    route: str
    county: str
    pp: str
    pm: str

    def as_text(self) -> str:
        return f"{self.route}|{self.county}|{self.pp or '<blank>'}{self.pm}"


@dataclasses.dataclass
class Record:
    values: dict[str, object]
    identity: Identity
    district: str
    route_suffix: str
    provenance: dict[str, object]


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _file_binding(path: Path) -> dict[str, object]:
    st = path.stat()
    return {
        "path": str(path), "size": st.st_size,
        "mtime_utc": dt.datetime.fromtimestamp(st.st_mtime, dt.timezone.utc).isoformat(),
        "sha256": _sha256(path),
    }


def _source_contract(bindings: dict[str, dict[str, object]]) -> dict[str, object]:
    issues: list[dict[str, object]] = []
    members: dict[str, dict[str, object]] = {}
    for role, expected in EXPECTED_SOURCES.items():
        actual = bindings[role]
        ok = actual["size"] == expected["size"] and actual["sha256"] == expected["sha256"]
        members[role] = {"expected": expected, "actual": {
            "size": actual["size"], "sha256": actual["sha256"]}, "pass": ok}
        if not ok:
            issues.append({"kind": "source_contract_mismatch", "role": role,
                           "expected": expected, "actual": members[role]["actual"]})
    return {"pass": not issues, "members": members, "issues": issues}


def _s(value: object) -> str:
    return "" if value is None else str(value).strip()


def _space(value: object) -> str:
    return " ".join(_s(value).split())


def _number(value: object) -> str:
    text = _s(value)
    if not text:
        return ""
    try:
        num = Decimal(text)
    except InvalidOperation:
        return text.upper()
    if num == 0:
        return "0"
    rendered = format(num.normalize(), "f")
    return rendered.rstrip("0").rstrip(".") if "." in rendered else rendered


def _location(value: object) -> tuple[str, str, str, str]:
    text = _space(value).upper()
    match = _LOC_RE.fullmatch(text)
    if not match:
        raise ValueError(f"unrecognized LOCATION {text!r}")
    return match.group(1), match.group(2), f"{int(match.group(3)):03d}", match.group(4)


def _identity(values: dict[str, object]) -> tuple[Identity, str, str]:
    district, county, route, suffix = _location(values["LOCATION"])
    ident = Identity(route, county, _s(values["PP"]).upper(), _number(values["POST_MILE"]))
    return ident, district, suffix


def _canonical(field: str, value: object) -> str:
    if field in NUMERIC_FIELDS:
        return _number(value)
    text = _space(value)
    return text.upper() if field != "DESCRIPTION" else text


def _words(chars: list[dict[str, object]]) -> list[dict[str, object]]:
    out: list[dict[str, object]] = []
    current: dict[str, object] | None = None
    last_x1: float | None = None
    for char in sorted(chars, key=lambda c: float(c["x0"])):
        x0, x1 = float(char["x0"]), float(char["x1"])
        if last_x1 is None or x0 - last_x1 > 1.6:
            if current:
                out.append(current)
            current = {"text": str(char["text"]), "x0": x0, "x1": x1}
        else:
            assert current is not None
            current["text"] = str(current["text"]) + str(char["text"])
            current["x1"] = x1
        last_x1 = x1
    if current:
        out.append(current)
    return out


def _assign(words: list[dict[str, object]], windows: tuple[tuple[str, int, int], ...]) -> dict[str, str]:
    found: dict[str, list[str]] = {name: [] for name, _lo, _hi in windows}
    for word in words:
        best: str | None = None
        overlap = 0.0
        for name, lo, hi in windows:
            candidate = min(float(word["x1"]), hi) - max(float(word["x0"]), lo)
            if candidate > overlap:
                best, overlap = name, candidate
        if best is not None:
            found[best].append(str(word["text"]))
    return {name: " ".join(parts).strip() for name, parts in found.items()}


def _clusters(chars: list[dict[str, object]], tolerance: float = 3.0):
    rows: list[list[dict[str, object]]] = []
    tops: list[float] = []
    for char in sorted(chars, key=lambda c: (float(c["top"]), float(c["x0"]))):
        top = float(char["top"])
        if rows and abs(top - tops[-1]) <= tolerance:
            rows[-1].append(char)
        else:
            rows.append([char])
            tops.append(top)
    return zip(tops, rows)


def load_xlsx(path: Path) -> tuple[list[Record], dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["Sheet 1"]:
            raise ValueError(f"unexpected workbook sheets: {workbook.sheetnames}")
        sheet = workbook["Sheet 1"]
        iterator = sheet.iter_rows(values_only=True)
        header = tuple(next(iterator, ()) or ())
        if header != XLSX_HEADER:
            raise ValueError(f"unexpected XLSX header: {header!r}")
        records: list[Record] = []
        for excel_row, row in enumerate(iterator, 2):
            if not row or all(value in (None, "") for value in row):
                continue
            values = dict(zip(XLSX_HEADER, row))
            ident, district, suffix = _identity(values)
            records.append(Record(values, ident, district, suffix, {"excel_row": excel_row}))
        metadata = {
            "sheet": sheet.title, "header": list(header), "rows": len(records),
            "columns": len(header),
            "created": str(workbook.properties.created),
            "modified": str(workbook.properties.modified),
        }
        return records, metadata
    finally:
        workbook.close()


def load_detail_pdf(path: Path) -> tuple[list[Record], dict[str, object], list[dict[str, object]]]:
    records: list[Record] = []
    issues: list[dict[str, object]] = []
    pending: tuple[int, float, dict[str, str]] | None = None
    with pdfplumber.open(path) as pdf:
        metadata = dict(pdf.metadata or {})
        pages = len(pdf.pages)
        page_text = [(pdf.pages[i].extract_text() or "") for i in range(min(2, pages))]
        for page_number, page in enumerate(pdf.pages, 1):
            chars = [c for c in page.chars if str(c.get("text", "")).strip()
                     and str(c.get("fontname", "")).endswith("Courier")]
            for top, line_chars in _clusters(chars):
                words = _words(line_chars)
                line1 = _assign(words, LINE1_WINDOWS)
                location_ok = _LOC_RE.fullmatch(_space(line1["LOCATION"]).upper())
                if _PM_RE.fullmatch(line1["POST_MILE"]) and location_ok:
                    if pending is not None:
                        issues.append({"kind": "line1_without_line2", "page": pending[0], "top": pending[1]})
                    pending = (page_number, top, line1)
                    continue
                if pending is None:
                    continue
                l1_page, l1_top, l1 = pending
                pending = None
                line2 = _assign(words, LINE2_WINDOWS)
                values = {name: l1[name] for name in LINE1_FIELDS}
                values.update({name: line2[name] for name in LINE2_FIELDS})
                try:
                    ident, district, suffix = _identity(values)
                except ValueError as exc:
                    issues.append({"kind": "bad_identity", "page": l1_page, "top": l1_top, "error": str(exc)})
                    continue
                extras = {name: line2[name] for name in ("PRINT_CROSS_ROUTE_SUFFIX", "PRINT_XING_ROUTE_SUFFIX")}
                if any(extras.values()):
                    issues.append({"kind": "unmapped_print_tail_value", "identity": ident.as_text(),
                                   "page": l1_page, "values": extras})
                records.append(Record(values, ident, district, suffix,
                                      {"line1_page": l1_page, "line1_top": round(l1_top, 3),
                                       "line2_page": page_number, "line2_top": round(top, 3)}))
            page.flush_cache()
            if page_number % 100 == 0:
                print(f"detail PDF: {page_number}/{pages} pages, {len(records)} records", file=sys.stderr)
        if pending is not None:
            issues.append({"kind": "final_line1_without_line2", "page": pending[0], "top": pending[1]})
    report_date = re.search(r"REPORT DATE\s*:\s*([^\n]+)", "\n".join(page_text), re.I)
    reference_date = re.search(r"REFERENCE DATE\s*:\s*([^\n]+)", "\n".join(page_text), re.I)
    event = re.search(r"EVENT ID\s*:\s*\n?\s*(\d+)", "\n".join(page_text), re.I)
    return records, {
        "pages": pages, "metadata": metadata,
        "report_date": report_date.group(1).strip() if report_date else None,
        "reference_date": reference_date.group(1).strip() if reference_date else None,
        "event_id": event.group(1) if event else None,
        "records": len(records),
    }, issues


def _field_relation(field: str, excel_value: object, pdf_value: object, description_limit: int) -> tuple[str, str | None]:
    raw_excel, raw_pdf = _s(excel_value), _s(pdf_value)
    if raw_excel == raw_pdf:
        return "exact", None
    # The signature marker is added by the PDF renderer.  Do not symmetrically
    # erase it from both sources: Excel-flagged/PDF-unflagged is a source change.
    if field in DATE_FIELDS:
        match = _PDF_FLAGGED_DATE_RE.fullmatch(_space(pdf_value))
        if match and match.group(1) == _space(excel_value):
            return "render_equivalent", "pdf_added_print_signature_flag"
    excel, rendered = _canonical(field, excel_value), _canonical(field, pdf_value)
    if excel == rendered:
        if field in NUMERIC_FIELDS:
            return "render_equivalent", "numeric_padding_or_type"
        return "render_equivalent", "whitespace_or_case"
    # Oracle Reports clips the fixed DESCRIPTION cell at its independently
    # measured 32-character boundary.  Require the exact capacity, the measured
    # run-wide limit, a strictly longer source, and an exact prefix.
    description_capacity = int(DESCRIPTION_PRINT_CONTRACT["characters"])
    # A space in character 32 is not returned as trailing extracted text, so
    # compare the PDF to the right-trimmed exact 32-character source slice.  As
    # DESCRIPTION canonicalization has already collapsed runs of whitespace,
    # that permits only lengths 32 or 31, never an arbitrary shorter prefix.
    boundary_render = excel[:description_capacity].rstrip()
    if (
        field == "DESCRIPTION"
        and description_limit == description_capacity
        and len(excel) > description_capacity
        and len(rendered) in {description_capacity - 1, description_capacity}
        and rendered == boundary_render
    ):
        return "render_equivalent", "fixed_width_description_truncation"
    # A one-character printed lane cell renders a multi-digit lane value as '*'.
    if field == "MAIN_NL" and rendered == "*" and re.fullmatch(r"\d{2,}", excel):
        return "render_equivalent", "single_cell_numeric_overflow"
    return "unresolved_mismatch", None


def _timing_contract(excel_meta: dict[str, object], detail_meta: dict[str, object]) -> dict[str, object]:
    xlsx_created = str(excel_meta.get("created"))
    pdf_created = str((detail_meta.get("metadata") or {}).get("CreationDate"))
    expected = (xlsx_created == EXPECTED_TIMING["detail_xlsx_created"]
                and pdf_created == EXPECTED_TIMING["detail_pdf_creation"])
    before = False
    try:
        xdt = dt.datetime.strptime(xlsx_created, "%Y-%m-%d %H:%M:%S")
        pdt = dt.datetime.strptime(pdf_created, "D:%Y%m%d%H%M%S")
        before = pdt < xdt
    except ValueError:
        pass
    return {
        "expected": EXPECTED_TIMING, "actual": {
            "detail_pdf_creation": pdf_created, "detail_xlsx_created": xlsx_created},
        "pdf_before_xlsx": before,
        "comparison_basis": "source-embedded naive local timestamps from the bound files",
        "pass": expected and before,
    }


def _authorized_source_delta(left: Record, right: Record, field: str,
                             relations: dict[str, str], timing: dict[str, object]) -> bool:
    identity = (left.identity.route, left.identity.county, left.identity.pp, left.identity.pm)
    candidate = (
        identity == BOUND_DETAIL_DELTA["identity"]
        and right.identity == left.identity
        and field == BOUND_DETAIL_DELTA["field"]
        and _s(left.values[field]) == BOUND_DETAIL_DELTA["excel"]
        and _s(right.values[field]) == BOUND_DETAIL_DELTA["pdf"]
    )
    other_cells_clean = all(
        relation in {"exact", "render_equivalent"}
        for other_field, relation in relations.items() if other_field != field
    )
    return bool(candidate and timing.get("pass") and other_cells_clean)


def _negative_mutation_self_check(timing: dict[str, object]) -> dict[str, object]:
    identity = Identity(*BOUND_DETAIL_DELTA["identity"])
    excel_values = {field: "" for field in XLSX_HEADER}
    pdf_values = dict(excel_values)
    excel_values["DESCRIPTION"] = BOUND_DETAIL_DELTA["excel"]
    pdf_values["DESCRIPTION"] = BOUND_DETAIL_DELTA["pdf"]
    left = Record(excel_values, identity, "07", "", {"excel_row": -1})
    right = Record(pdf_values, identity, "07", "", {"line1_page": -1})
    baseline_relations = {
        field: _field_relation(field, left.values[field], right.values[field], 32)[0]
        for field in XLSX_HEADER
    }
    bound_alone_authorized = _authorized_source_delta(
        left, right, "DESCRIPTION", baseline_relations, timing)

    # Mutate a second cell on the otherwise permitted VEN pair.  It must be
    # unresolved and must revoke authorization of the DESCRIPTION delta.
    right.values["HG"] = "U"
    left.values["HG"] = "D"
    mutated_relations = {
        field: _field_relation(field, left.values[field], right.values[field], 32)[0]
        for field in XLSX_HEADER
    }
    second_relation = mutated_relations["HG"]
    after_second_authorized = _authorized_source_delta(
        left, right, "DESCRIPTION", mutated_relations, timing)
    summary_mutation_relation = _summary_relation(1, 2)
    capacity = int(DESCRIPTION_PRINT_CONTRACT["characters"])
    pdf_added_flag_relation = _field_relation(
        "EFF_DATE_LT", "01-02-03", "Y01-02-03", capacity)[0]
    reverse_flag_relation = _field_relation(
        "EFF_DATE_LT", "Y01-02-03", "01-02-03", capacity)[0]
    exact_boundary_truncation_relation = _field_relation(
        "DESCRIPTION", "A" * (capacity + 1), "A" * capacity, capacity)[0]
    space_at_boundary_truncation_relation = _field_relation(
        "DESCRIPTION", "A" * (capacity - 1) + " B", "A" * (capacity - 1), capacity)[0]
    short_prefix_relation = _field_relation(
        "DESCRIPTION", "A" * (capacity + 5), "A" * (capacity - 2), capacity)[0]
    wrong_measured_limit_relation = _field_relation(
        "DESCRIPTION", "A" * (capacity + 1), "A" * capacity, capacity + 1)[0]
    passed = (bound_alone_authorized and second_relation == "unresolved_mismatch"
              and not after_second_authorized
              and summary_mutation_relation == "unresolved_mismatch"
              and pdf_added_flag_relation == "render_equivalent"
              and reverse_flag_relation == "unresolved_mismatch"
              and exact_boundary_truncation_relation == "render_equivalent"
              and space_at_boundary_truncation_relation == "render_equivalent"
              and short_prefix_relation == "unresolved_mismatch"
              and wrong_measured_limit_relation == "unresolved_mismatch")
    return {
        "mutation": (
            "VEN bound pair plus changed HG; directional date-flag probes; "
            "exact and inexact DESCRIPTION-boundary probes"
        ),
        "bound_delta_authorized_before_second_change": bound_alone_authorized,
        "second_changed_cell_relation": second_relation,
        "bound_delta_authorized_after_second_change": after_second_authorized,
        "changed_summary_count_relation": summary_mutation_relation,
        "pdf_added_date_flag_relation": pdf_added_flag_relation,
        "reverse_direction_date_flag_relation": reverse_flag_relation,
        "exact_32_character_boundary_relation": exact_boundary_truncation_relation,
        "space_in_character_32_boundary_relation": space_at_boundary_truncation_relation,
        "short_shared_prefix_relation": short_prefix_relation,
        "wrong_measured_limit_relation": wrong_measured_limit_relation,
        "pass": passed,
    }


def _pair_group(excel: list[Record], pdf: list[Record], description_limit: int):
    if len(excel) != len(pdf):
        return None, "multiplicity_mismatch"
    n = len(excel)
    if n == 1:
        return [(excel[0], pdf[0])], None
    if n > 8:
        return None, "duplicate_group_too_large_for_unambiguous_assignment"
    matrix: list[list[int]] = []
    for left in excel:
        costs: list[int] = []
        for right in pdf:
            costs.append(sum(_field_relation(field, left.values[field], right.values[field], description_limit)[0]
                             == "unresolved_mismatch" for field in XLSX_HEADER))
        matrix.append(costs)
    candidates: list[tuple[int, tuple[int, ...]]] = []
    for permutation in itertools.permutations(range(n)):
        candidates.append((sum(matrix[i][j] for i, j in enumerate(permutation)), permutation))
    best_cost = min(cost for cost, _permutation in candidates)
    best = [permutation for cost, permutation in candidates if cost == best_cost]
    if len(best) > 1:
        excel_fps = {tuple(_canonical(f, r.values[f]) for f in XLSX_HEADER) for r in excel}
        pdf_fps = {tuple(_canonical(f, r.values[f]) for f in XLSX_HEADER) for r in pdf}
        if len(excel_fps) == len(pdf_fps) == 1:
            return list(zip(excel, pdf)), "interchangeable_exact_full_duplicates"
        return None, f"ambiguous_minimum_cost_assignment:{len(best)}"
    return [(excel[i], pdf[j]) for i, j in enumerate(best[0])], None


def detail_parity(excel_records: list[Record], pdf_records: list[Record],
                  parser_issues: list[dict[str, object]], timing: dict[str, object]):
    excel_by: dict[Identity, list[Record]] = collections.defaultdict(list)
    pdf_by: dict[Identity, list[Record]] = collections.defaultdict(list)
    for record in excel_records:
        excel_by[record.identity].append(record)
    for record in pdf_records:
        pdf_by[record.identity].append(record)

    description_limit = max((len(_canonical("DESCRIPTION", r.values["DESCRIPTION"])) for r in pdf_records), default=0)
    unmatched_excel = sorted(set(excel_by) - set(pdf_by), key=lambda key: key.as_text())
    unmatched_pdf = sorted(set(pdf_by) - set(excel_by), key=lambda key: key.as_text())
    unresolved = list(parser_issues)
    if description_limit != DESCRIPTION_PRINT_CONTRACT["characters"]:
        unresolved.append({
            "kind": "description_print_capacity_contract_failed",
            "expected": DESCRIPTION_PRINT_CONTRACT,
            "actual_maximum_characters": description_limit,
        })
    pairs: list[tuple[Record, Record]] = []
    duplicate_groups = 0
    interchangeable_groups = 0
    for key in sorted(set(excel_by) & set(pdf_by), key=lambda item: item.as_text()):
        left, right = excel_by[key], pdf_by[key]
        if len(left) > 1 or len(right) > 1:
            duplicate_groups += 1
        matched, issue = _pair_group(left, right, description_limit)
        if matched is None:
            unresolved.append({"kind": issue, "identity": key.as_text(),
                               "excel_count": len(left), "pdf_count": len(right)})
            continue
        if issue == "interchangeable_exact_full_duplicates":
            interchangeable_groups += 1
        pairs.extend(matched)

    relation_counts = collections.Counter()
    render_reasons = collections.Counter()
    field_counts: dict[str, collections.Counter] = {field: collections.Counter() for field in XLSX_HEADER}
    delta_cells: list[dict[str, object]] = []
    suffix_mismatches: list[dict[str, object]] = []
    for left, right in pairs:
        if left.route_suffix != right.route_suffix:
            suffix_mismatches.append({"identity": left.identity.as_text(), "excel": left.route_suffix,
                                      "pdf": right.route_suffix, "excel_row": left.provenance["excel_row"],
                                      "pdf_page": right.provenance["line1_page"]})
        evaluated = {
            field: _field_relation(field, left.values[field], right.values[field], description_limit)
            for field in XLSX_HEADER
        }
        raw_relations = {field: relation for field, (relation, _reason) in evaluated.items()}
        for field in XLSX_HEADER:
            relation, reason = evaluated[field]
            if relation == "unresolved_mismatch" and _authorized_source_delta(
                    left, right, field, raw_relations, timing):
                relation = "source_export_delta"
            relation_counts[relation] += 1
            field_counts[field][relation] += 1
            if reason:
                render_reasons[(field, reason)] += 1
            if relation == "source_export_delta":
                delta_cells.append({
                    "identity": left.identity.as_text(), "field": field,
                    "excel": _s(left.values[field]), "pdf": _s(right.values[field]),
                    "excel_row": left.provenance["excel_row"],
                    "pdf_page": right.provenance["line1_page"],
                })
            elif relation == "unresolved_mismatch":
                unresolved.append({
                    "kind": "unapproved_detail_cell_mismatch",
                    "identity": left.identity.as_text(), "field": field,
                    "excel": _s(left.values[field]), "pdf": _s(right.values[field]),
                    "excel_row": left.provenance["excel_row"],
                    "pdf_page": right.provenance["line1_page"],
                })

    full_row_counts = collections.Counter(tuple(_s(r.values[f]) for f in XLSX_HEADER) for r in excel_records)
    exact_duplicate_groups = sum(1 for count in full_row_counts.values() if count > 1)
    exact_duplicate_rows = sum(count for count in full_row_counts.values() if count > 1)

    weak: dict[tuple[str, str, str], set[str]] = collections.defaultdict(set)
    for record in excel_records:
        weak[(record.identity.route, record.identity.county, record.identity.pm)].add(record.identity.pp)
    actual_prefix_pairs = {key: prefixes for key, prefixes in weak.items() if len(prefixes) > 1}
    prefix_canaries = []
    for key, expected in EXPECTED_PREFIX_PAIRS.items():
        actual = actual_prefix_pairs.get(key, set())
        prefix_canaries.append({"route": key[0], "county": key[1], "pm": key[2],
                                "expected_pp": sorted(expected), "actual_pp": sorted(actual),
                                "pass": actual == expected})
        if actual != expected:
            unresolved.append({"kind": "prefix_canary_failed", "key": list(key),
                               "expected": sorted(expected), "actual": sorted(actual)})

    if unmatched_excel:
        unresolved.append({"kind": "xlsx_identities_absent_from_pdf", "count": len(unmatched_excel),
                           "identities": [key.as_text() for key in unmatched_excel]})
    if unmatched_pdf:
        unresolved.append({"kind": "pdf_identities_absent_from_xlsx", "count": len(unmatched_pdf),
                           "identities": [key.as_text() for key in unmatched_pdf]})
    if suffix_mismatches:
        unresolved.append({"kind": "route_suffix_claim_mismatch", "count": len(suffix_mismatches),
                           "cells": suffix_mismatches})
    expected_delta = {
        "identity": "001|VEN|<blank>23.907", "field": BOUND_DETAIL_DELTA["field"],
        "excel": BOUND_DETAIL_DELTA["excel"], "pdf": BOUND_DETAIL_DELTA["pdf"],
    }
    observed_delta_core = [{key: cell[key] for key in expected_delta} for cell in delta_cells]
    if observed_delta_core != [expected_delta]:
        unresolved.append({"kind": "bound_detail_delta_contract_failed",
                           "expected": [expected_delta], "observed": observed_delta_core})

    return {
        "xlsx_records": len(excel_records), "pdf_records": len(pdf_records),
        "paired_records": len(pairs), "asserted_cells": len(pairs) * len(XLSX_HEADER),
        "description_print_limit": description_limit,
        "description_print_contract": DESCRIPTION_PRINT_CONTRACT,
        "relation_counts": dict(relation_counts),
        "render_equivalence": [{"field": field, "reason": reason, "count": count}
                               for (field, reason), count in sorted(render_reasons.items())],
        "per_field": {field: dict(field_counts[field]) for field in XLSX_HEADER},
        "source_export_delta_count": len(delta_cells),
        "source_export_deltas": delta_cells,
        "source_export_delta_contract": {
            "allowlist": BOUND_DETAIL_DELTA,
            "requires_other_35_cells_clean": True,
            "timing": timing,
            "pass": observed_delta_core == [expected_delta],
        },
        "physical_identity": {
            "definition": ["base_route", "county", "complete_PP", "numeric_POST_MILE"],
            "route_suffix_policy": "preserved and asserted separately; never discarded",
            "xlsx_unique_identities": len(excel_by), "pdf_unique_identities": len(pdf_by),
            "prefix_collision_groups": len(actual_prefix_pairs),
            "six_prefix_canaries": prefix_canaries,
        },
        "duplicates": {
            "physical_duplicate_groups_seen": duplicate_groups,
            "interchangeable_full_duplicate_groups_paired_as_multisets": interchangeable_groups,
            "xlsx_exact_full_duplicate_groups": exact_duplicate_groups,
            "xlsx_rows_in_exact_full_duplicate_groups": exact_duplicate_rows,
        },
        "unresolved": unresolved,
    }


def _summary_section(text: str) -> str | None:
    normalized = re.sub(r"[^A-Z0-9/+ ]+", " ", text.upper())
    normalized = " ".join(normalized.split())
    for section in SUMMARY_CODES:
        if section in normalized:
            return section
    return None


def _summary_code(section: str, rest: list[str], parent: str | None) -> tuple[str | None, str | None]:
    if not rest:
        return None, parent
    token = rest[0].upper()
    if section == "MAINLINE NUM OF LANES":
        code = "+" if token.startswith("+") else token
        return code if code in SUMMARY_CODES[section] else None, parent
    if section == "RURAL/URBAN/SUBURBAN":
        if token.startswith("R-"):
            return "R", "R"
        if token.startswith("U-"):
            return "U", "U"
        if token.startswith("-O") and parent in {"R", "U"}:
            return f"{parent}-O", parent
        if token.startswith("+"):
            return "+", parent
        return None, parent
    code = token.split("-", 1)[0]
    if code == "":
        return None, parent
    if code.startswith("+"):
        code = "+"
    return (code if code in SUMMARY_CODES[section] else None), parent


def load_summary_pdf(path: Path):
    with pdfplumber.open(path) as pdf:
        page = next((pg for pg in pdf.pages if "Total Intersections" in (pg.extract_text() or "")), None)
        if page is None:
            raise ValueError("summary total page not found")
        text = page.extract_text() or ""
        total_match = re.search(r"Total Intersections\s*=\s*([\d,]+)", text, re.I)
        total = int(total_match.group(1).replace(",", "")) if total_match else None
        bands = [(0, 190), (190, 495), (495, float(page.width) + 1)]
        parsed: dict[str, dict[str, int]] = collections.defaultdict(dict)
        raw_rows: list[dict[str, object]] = []
        words = page.extract_words()
        for band_number, (left, right) in enumerate(bands, 1):
            band_words = [w for w in words if left <= float(w["x0"]) < right]
            rows: list[list[dict[str, object]]] = []
            for word in sorted(band_words, key=lambda w: (float(w["top"]), float(w["x0"]))):
                if rows and abs(float(word["top"]) - float(rows[-1][0]["top"])) <= 3:
                    rows[-1].append(word)
                else:
                    rows.append([word])
            section = None
            rural_parent = None
            for row in rows:
                row.sort(key=lambda w: float(w["x0"]))
                tokens = [str(w["text"]) for w in row]
                line = " ".join(tokens)
                heading = _summary_section(line)
                if heading:
                    section = heading
                    rural_parent = None
                    continue
                if section is None or not tokens or not re.fullmatch(r"-?[\d,]+", tokens[0]):
                    continue
                count = int(tokens[0].replace(",", ""))
                code, rural_parent = _summary_code(section, tokens[1:], rural_parent)
                if code is None:
                    continue
                if code in parsed[section]:
                    raise ValueError(f"duplicate summary category {section}/{code}")
                parsed[section][code] = count
                raw_rows.append({"band": band_number, "section": section, "code": code,
                                 "count": count, "text": " ".join(tokens[1:])})
        metadata = dict(pdf.metadata or {})
        pages = len(pdf.pages)
    return dict(parsed), total, {"pages": pages, "metadata": metadata, "raw_category_rows": raw_rows}


def _cat(value: object, allowed: tuple[str, ...]) -> str:
    code = _s(value).upper()
    # Blank is the report's explicit '+ no data' bucket.  A nonblank code outside
    # the printed taxonomy must remain itself: folding e.g. Roundabout 'R' into '+'
    # would manufacture a false summary discrepancy and lose the source fact that
    # the Summary print simply does not tabulate that code.
    return "+" if not code else code


def aggregate_summary(records: list[Record]):
    counts: dict[str, collections.Counter] = {section: collections.Counter() for section in SUMMARY_CODES}
    members: dict[tuple[str, str], list[dict[str, object]]] = collections.defaultdict(list)
    residue: list[dict[str, object]] = []
    taxonomy_excluded: list[dict[str, object]] = []
    for record in records:
        values = record.values
        section_values: dict[str, str] = {}
        section_values["HIGHWAY GROUP"] = _cat(values["HG"], SUMMARY_CODES["HIGHWAY GROUP"])
        ru = _s(values["RU"]).upper()
        inside = bool(_s(values["CITY_CODE"]))
        if ru == "R":
            section_values["RURAL/URBAN/SUBURBAN"] = "R" if inside else "R-O"
        elif ru in {"U", "B"}:
            section_values["RURAL/URBAN/SUBURBAN"] = "U" if inside else "U-O"
        else:
            section_values["RURAL/URBAN/SUBURBAN"] = "+"
        section_values["INTERSECTION TYPE"] = _cat(values["TY_INT"], SUMMARY_CODES["INTERSECTION TYPE"])
        section_values["LIGHTING TYPE"] = _cat(values["LT_TY"], SUMMARY_CODES["LIGHTING TYPE"])
        section_values["CONTROL TYPES"] = _cat(values["TY_CT"], SUMMARY_CODES["CONTROL TYPES"])
        section_values["MAINLINE NUM OF LANES"] = _cat(_number(values["MAIN_NL"]), SUMMARY_CODES["MAINLINE NUM OF LANES"])
        section_values["MAINLINE MASTARM"] = _cat(values["MAIN_SM"], SUMMARY_CODES["MAINLINE MASTARM"])
        section_values["MAINLINE LEFT CHANNELIZATION"] = _cat(values["MAIN_LC"], SUMMARY_CODES["MAINLINE LEFT CHANNELIZATION"])
        section_values["MAINLINE RIGHT CHANNELIZATION"] = _cat(values["MAIN_RC"], SUMMARY_CODES["MAINLINE RIGHT CHANNELIZATION"])
        section_values["MAINLINE TRAFFIC FLOW"] = _cat(values["MAIN_TF"], SUMMARY_CODES["MAINLINE TRAFFIC FLOW"])
        for section, code in section_values.items():
            counts[section][code] += 1
            raw = {
                "HIGHWAY GROUP": values["HG"], "RURAL/URBAN/SUBURBAN": values["RU"],
                "INTERSECTION TYPE": values["TY_INT"], "LIGHTING TYPE": values["LT_TY"],
                "CONTROL TYPES": values["TY_CT"], "MAINLINE NUM OF LANES": values["MAIN_NL"],
                "MAINLINE MASTARM": values["MAIN_SM"],
                "MAINLINE LEFT CHANNELIZATION": values["MAIN_LC"],
                "MAINLINE RIGHT CHANNELIZATION": values["MAIN_RC"],
                "MAINLINE TRAFFIC FLOW": values["MAIN_TF"],
            }[section]
            members[(section, code)].append({
                "identity": record.identity.as_text(),
                "excel_row": record.provenance.get("excel_row"), "raw": _s(raw),
            })
            if code not in SUMMARY_CODES[section]:
                taxonomy_excluded.append({
                    "identity": record.identity.as_text(), "section": section,
                    "raw": _s(raw), "excel_row": record.provenance.get("excel_row"),
                    "reason": "the TSN Summary print has no category for this nonblank detail code",
                })
    return ({section: dict(counter) for section, counter in counts.items()},
            residue, taxonomy_excluded, members)


def _summary_relation(excel: int, printed: int | None) -> str:
    return "exact" if printed == excel else "unresolved_mismatch"


def summary_parity(excel_records: list[Record], pdf_counts: dict[str, dict[str, int]], total: int | None,
                   pdf_meta: dict[str, object]):
    excel_counts, residue, taxonomy_excluded, members = aggregate_summary(excel_records)
    cells: list[dict[str, object]] = []
    unresolved: list[dict[str, object]] = []
    for section, codes in SUMMARY_CODES.items():
        missing = sorted(set(codes) - set(pdf_counts.get(section, {})))
        extra = sorted(set(pdf_counts.get(section, {})) - set(codes))
        if missing or extra:
            unresolved.append({"kind": "summary_taxonomy_loss", "section": section,
                               "missing_pdf_codes": missing, "extra_pdf_codes": extra})
        for code in codes:
            excel = excel_counts.get(section, {}).get(code, 0)
            printed = pdf_counts.get(section, {}).get(code)
            relation = _summary_relation(excel, printed)
            cells.append({"section": section, "code": code, "excel": excel,
                          "pdf": printed, "relation": relation})
            if relation != "exact":
                unresolved.append({"kind": "unapproved_summary_category_mismatch",
                                   "section": section, "code": code,
                                   "excel": excel, "pdf": printed})
    section_conservation = []
    for section in SUMMARY_CODES:
        excel_sum = sum(excel_counts.get(section, {}).get(code, 0) for code in SUMMARY_CODES[section])
        pdf_sum = sum(pdf_counts.get(section, {}).values())
        excluded = sum(1 for row in taxonomy_excluded if row["section"] == section)
        section_conservation.append({"section": section, "excel_sum": excel_sum,
                                     "pdf_sum": pdf_sum, "expected_total": len(excel_records),
                                     "xlsx_rows_excluded_by_print_taxonomy": excluded,
                                     "xlsx_rows_accounted": excel_sum + excluded,
                                     "pass": excel_sum + excluded == len(excel_records)})
        if excel_sum + excluded != len(excel_records):
            unresolved.append({"kind": "summary_xlsx_rows_not_accounted", "section": section,
                               "mapped_excel_sum": excel_sum, "excluded": excluded})
    if total != len(excel_records):
        unresolved.append({"kind": "summary_total_mismatch", "xlsx_rows": len(excel_records), "pdf_total": total})
    if residue:
        unresolved.append({"kind": "unmapped_excel_summary_values", "count": len(residue),
                           "rows": residue})

    folded_excel: dict[tuple[str, str], int] = {}
    folded_pdf: dict[tuple[str, str], int] = {}
    for section, codes in SUMMARY_CODES.items():
        for code in codes:
            out_code = "S" if section == "CONTROL TYPES" and code in SIGNAL_CONTROL_CODES else code
            folded_excel[(section, out_code)] = folded_excel.get((section, out_code), 0) + excel_counts[section].get(code, 0)
            folded_pdf[(section, out_code)] = folded_pdf.get((section, out_code), 0) + pdf_counts.get(section, {}).get(code, 0)
    folded_cells = [{"section": section, "code": code, "excel": count,
                     "pdf": folded_pdf.get((section, code)),
                     "relation": _summary_relation(count, folded_pdf.get((section, code)))}
                    for (section, code), count in sorted(folded_excel.items())]
    normalized_count = len(folded_cells) + 1  # + statewide total
    if normalized_count != 58:
        unresolved.append({"kind": "normalized_summary_category_count", "expected": 58,
                           "actual": normalized_count})
    mismatch_members = [{"section": cell["section"], "code": cell["code"],
                         "excel": cell["excel"], "pdf": cell["pdf"],
                         "xlsx_rows": members.get((cell["section"], cell["code"]), [])}
                        for cell in cells if cell["relation"] != "exact"]
    return {
        "pdf_metadata": pdf_meta, "pdf_total": total,
        "raw_pdf_category_count": sum(len(v) for v in pdf_counts.values()),
        "raw_category_cells": cells,
        "raw_category_delta_count": sum(cell["relation"] != "exact" for cell in cells),
        "source_export_delta_members": [],
        "unresolved_mismatch_members": mismatch_members,
        "section_conservation": section_conservation,
        "fold_rule": "CONTROL TYPES J/K/L/M/N/P -> S (Signalized)",
        "normalized_categories_including_total": normalized_count,
        "normalized_cells": folded_cells,
        "normalized_delta_count": sum(cell["relation"] != "exact" for cell in folded_cells),
        "unmapped_excel_category_values": residue,
        "xlsx_rows_excluded_by_print_taxonomy": taxonomy_excluded,
        "unresolved": unresolved,
    }


def report_view_mapping() -> dict[str, object]:
    line1 = [{"pdf_line": 1, "pdf_cell": name, "xlsx_column": name,
              "report_view_role": ("identity" if name in {"PP", "POST_MILE", "LOCATION"} else "mainline")}
             for name in LINE1_FIELDS]
    line2 = [{"pdf_line": 2, "pdf_cell": name, "xlsx_column": name,
              "report_view_role": ("cross_street" if name not in {"DESCRIPTION", "MAIN_OVERRIDE"}
                                   else "description_or_mainline_length")}
             for name in LINE2_FIELDS]
    return {
        "record_shape": "two physical PDF lines per one Excel row",
        "line_1_fields": line1, "line_2_fields": line2,
        "all_xlsx_columns_mapped_once": sorted(x["xlsx_column"] for x in line1 + line2) == sorted(XLSX_HEADER),
        "print_only_blank_cells": ["PRINT_CROSS_ROUTE_SUFFIX", "PRINT_XING_ROUTE_SUFFIX"],
    }


def _write_result(path: Path, result: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    temp.replace(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--detail-pdf", type=Path, default=DEFAULT_DETAIL_PDF)
    parser.add_argument("--summary-pdf", type=Path, default=DEFAULT_SUMMARY_PDF)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    for path in (args.xlsx, args.detail_pdf, args.summary_pdf):
        if not path.is_file():
            parser.error(f"source does not exist: {path}")

    bindings = {
        "detail_xlsx": _file_binding(args.xlsx),
        "detail_pdf": _file_binding(args.detail_pdf),
        "summary_pdf": _file_binding(args.summary_pdf),
    }
    source_contract = _source_contract(bindings)
    if not source_contract["pass"]:
        result = {
            "oracle": "phase4_intersection_tsn_pdf_oracle",
            "method": "independent, hash-pinned, fail-closed",
            "generated_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
            "status": "fail", "unresolved_gate_count": len(source_contract["issues"]),
            "sources": bindings, "source_contract": source_contract,
        }
        _write_result(args.output, result)
        print(json.dumps({"status": "fail", "output": str(args.output),
                          "source_contract_issues": source_contract["issues"]}, indent=2))
        return 1

    excel_records, excel_meta = load_xlsx(args.xlsx)
    pdf_records, detail_meta, parser_issues = load_detail_pdf(args.detail_pdf)
    timing = _timing_contract(excel_meta, detail_meta)
    detail = detail_parity(excel_records, pdf_records, parser_issues, timing)
    if not timing["pass"]:
        detail["unresolved"].append({"kind": "source_timing_contract_failed",
                                     "contract": timing})
    summary_counts, summary_total, summary_meta = load_summary_pdf(args.summary_pdf)
    summary = summary_parity(excel_records, summary_counts, summary_total, summary_meta)
    mapping = report_view_mapping()
    negative_check = _negative_mutation_self_check(timing)
    auxiliary_failures = int(not mapping["all_xlsx_columns_mapped_once"]) + int(not negative_check["pass"])
    unresolved_count = len(detail["unresolved"]) + len(summary["unresolved"]) + auxiliary_failures
    status = "pass" if unresolved_count == 0 else "fail"
    result = {
        "oracle": "phase4_intersection_tsn_pdf_oracle",
        "method": "independent, hash-pinned, fail-closed; no production comparator/parser/schema imports",
        "generated_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "status": status, "unresolved_gate_count": unresolved_count,
        "sources": {
            "detail_xlsx": {**bindings["detail_xlsx"], **excel_meta},
            "detail_pdf": {**bindings["detail_pdf"], **detail_meta},
            "summary_pdf": {**bindings["summary_pdf"], **summary_meta},
        },
        "source_contract": source_contract,
        "source_timing_contract": timing,
        "internal_negative_mutation_self_check": negative_check,
        "detail_cross_format": detail,
        "summary_from_detail": summary,
        "report_view_source_mapping": mapping,
    }
    _write_result(args.output, result)
    print(json.dumps({
        "status": status, "output": str(args.output),
        "xlsx_records": len(excel_records), "pdf_records": len(pdf_records),
        "detail_source_deltas": detail["source_export_delta_count"],
        "detail_unresolved": len(detail["unresolved"]),
        "summary_raw_deltas": summary["raw_category_delta_count"],
        "summary_unresolved": len(summary["unresolved"]),
    }, indent=2))
    return 0 if status == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
