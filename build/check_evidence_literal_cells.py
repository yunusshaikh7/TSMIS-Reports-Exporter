"""Golden check: evidence workbooks never execute source text as Excel code.

CMP-AUD-111: visual-evidence summaries and image-sheet captions contain values
parsed from source reports.  Formula-leading text and Excel's seven error tokens
must survive byte-for-byte as string cells, never formulas/errors.

Run with the build venv:
    build\.venv\Scripts\python.exe build\check_evidence_literal_cells.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import visual_evidence as ve
from compare_core import is_formula_injection
from openpyxl import load_workbook
from openpyxl.cell.cell import ERROR_CODES
from PIL import Image


FORMULA_LEADS = ("=1+1", "+1", "-1", "@SUM(A1:A2)")
UNSAFE_LITERALS = FORMULA_LEADS + tuple(ERROR_CODES)


def _assert_text(cell, expected, label):
    assert cell.value == expected, (label, cell.coordinate, cell.value, expected)
    assert cell.data_type == "s", (
        label, cell.coordinate, cell.value, cell.data_type,
    )


def test_evidence_workbook_literals():
    with tempfile.TemporaryDirectory(prefix="tsmis_evidence_literals_") as td:
        root = Path(td)
        image = root / "sample.png"
        Image.new("RGB", (8, 8), "white").save(image)

        entries = [
            {
                "field": value,
                "route": value,
                "key": "K",
                "va": value,
                "vb": value,
                "stacked": image.name,
                "pair": image.name,
                "note": value,
            }
            for value in UNSAFE_LITERALS
        ]
        out = root / "evidence.xlsx"
        note = ve._write_workbook(
            out,
            root,
            entries,
            {"=missing field": "@source unavailable"},
            {
                "report": "=report label",
                "comparison": "@comparison name",
                "examples": len(entries),
                "seed": "-seed",
                "tsmis_dir": "+tsmis path",
                "tsn_dir": "=tsn path",
            },
            layout="both",           # exercise BOTH per-column tab writers
        )
        assert note is None, note

        wb = load_workbook(out, data_only=False)
        summary = wb["Summary"]
        _assert_text(summary["A1"], "=report label — visual evidence",
                     "summary title")

        for offset, value in enumerate(UNSAFE_LITERALS, start=6):
            _assert_text(summary.cell(offset, 1), value, "summary field")
            _assert_text(summary.cell(offset, 2), f"{value} @ K",
                         "summary route/key")
            _assert_text(summary.cell(offset, 3), value, "summary TSMIS value")
            _assert_text(summary.cell(offset, 4), value, "summary TSN value")

        miss_row = 6 + len(UNSAFE_LITERALS)
        _assert_text(summary.cell(miss_row, 1), "=missing field",
                     "summary missing-field caption")

        # 'both' layout -> one tab PER COMPARISON COLUMN for each of the two
        # layouts. Each entry has a distinct formula-leading `field`, so every
        # field becomes its own (sanitized) tab: header row 1 leads with the
        # source `field` (formula-leading), captions start at row 4. EVERY
        # source-derived col-1 cell must remain a literal string.
        image_sheets = [s for s in wb.sheetnames if s != "Summary"]
        assert len(image_sheets) == 2 * len(UNSAFE_LITERALS), (
            len(image_sheets), 2 * len(UNSAFE_LITERALS))
        headers, captions = [], []
        for name in image_sheets:
            ws = wb[name]
            for (cell,) in ws.iter_rows(min_row=1, min_col=1, max_col=1):
                if cell.value is not None:            # legend row 2 is fixed text
                    assert cell.data_type == "s", (
                        name, cell.coordinate, cell.value, cell.data_type)
            headers.append(ws.cell(1, 1).value)
            captions.append(ws.cell(4, 1).value)
        for value in UNSAFE_LITERALS:
            caption = (f"route {value} @ K   —   TSMIS '{value}' vs "
                       f"TSN '{value}'   —   {value}")
            assert captions.count(caption) == 2, (value, captions.count(caption))
            assert sum(1 for h in headers if h.startswith(f"{value} — ")) == 2, (
                value, [h for h in headers if h.startswith(str(value))])
        wb.close()


def test_shared_guard_covers_excel_error_tokens():
    assert all(is_formula_injection(value) for value in UNSAFE_LITERALS)
    assert not any(is_formula_injection(value) for value in
                   ("normal", "#N/A with context", "1", 1, None))


def main():
    test_evidence_workbook_literals()
    test_shared_guard_covers_excel_error_tokens()
    print(
        "OK  EVIDENCE-LITERAL-CELLS: =+-@ source text and all openpyxl "
        "ERROR_CODES remain byte-exact string cells in summaries and captions."
    )


if __name__ == "__main__":
    main()
