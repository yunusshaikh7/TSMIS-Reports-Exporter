"""Golden check for the consolidate-overwrite TOCTOU re-check (P12).

The truncation half of `consolidate-overwrite-toctou` is already closed (the
producer writes a temp sibling, then an atomic os.replace — F9/P2, so a prior good
workbook is never truncated). The remaining half is the CONFIRM-THEN-APPEARS
window: the overwrite prompt is shown (or skipped) based on the destination's
existence at the START, but the file can APPEAR while the producer runs, and the old
code then ran os.replace and silently clobbered a file the user never agreed to
overwrite.

The gate (`artifact_store.confirm_late_overwrite`) is enforced at the FINAL commit
point — the os.replace inside `atomic_save_if` — for EVERY confirm->write path:
  * artifact_store.commit_workbook  (every compare_core comparator)
  * consolidate_xlsx_base.consolidate_xlsx  (highway log / sequence / intersection
    detail — via atomic_save_if)
  * the direct writers (ramp_summary, intersection_summary, tsn_highway_sequence) —
    each routes its build_workbook/_write_workbook save through atomic_save_if(proceed=...)
  * the per-route converters (tsmis/tsn highway_log) — pass the REAL confirm +
    existed_at_confirm into consolidate_xlsx so its gate catches a late appearance.

P12-B01 (Codex round 1): the earlier re-check sat BEFORE build_workbook/_write_workbook/
consolidate_xlsx, so a destination that appeared DURING the build (after the re-check,
before the actual os.replace) was still overwritten. These checks now make the
destination appear RIGHT AT the final save (by patching atomic_save_if to create it
just before the real gate runs) and prove a decline returns `cancelled`, preserves the
appeared file, and records a confirm call — for all four direct/converter shapes.
compare_core itself is untouched.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_consolidate_toctou.py
"""
import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import artifact_store
import consolidate_xlsx_base
import consolidate_ramp_summary
import consolidate_intersection_summary
import consolidate_tsn_highway_sequence
import consolidate_tsmis_highway_log_pdf
from events import ConsolidateResult, Events

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _tmp(prefix):
    return Path(tempfile.mkdtemp(prefix=prefix))


def _write_valid_xlsx(path, sheet="Sheet1", header=("A",), rows=(("1",),)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def _run_with_late_save_appearance(out_path, run):
    """Patch artifact_store.atomic_save_if so `out_path` APPEARS right as the final
    save begins (after the build, just before the os.replace gate), then run the
    consolidator. The consolidator's real gate (proceed=confirm_late_overwrite) must
    catch it. Restores the real atomic_save_if afterward."""
    real = artifact_store.atomic_save_if

    def inject(wb, out, proceed):
        Path(out).write_text("APPEARED", encoding="utf-8")   # external appearance, post-build
        return real(wb, out, proceed)

    artifact_store.atomic_save_if = inject
    try:
        return run()
    finally:
        artifact_store.atomic_save_if = real


# ----------------------------------------------------------------------------- #
# unit: confirm_late_overwrite
# ----------------------------------------------------------------------------- #
def unit_checks():
    print("confirm_late_overwrite (unit):")
    tmp = _tmp("tsmis_toctou_unit_")
    try:
        present = tmp / "present.xlsx"
        present.write_text("x", encoding="utf-8")
        absent = tmp / "absent.xlsx"
        calls = []

        def confirm(p):
            calls.append(p)
            return False

        check("existed-at-confirm -> proceed, confirm NOT re-called",
              artifact_store.confirm_late_overwrite(present, True, confirm) is True
              and not calls)
        check("absent-then-still-absent -> proceed, confirm NOT called",
              artifact_store.confirm_late_overwrite(absent, False, confirm) is True
              and not calls)
        check("appeared -> confirm IS called, decline returns False",
              artifact_store.confirm_late_overwrite(present, False, confirm) is False
              and len(calls) == 1)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ----------------------------------------------------------------------------- #
# atomic_save_if: the gate IS at the os.replace
# ----------------------------------------------------------------------------- #
def atomic_save_if_checks():
    print("atomic_save_if (the gate at the final os.replace):")
    from openpyxl import Workbook
    tmp = _tmp("tsmis_toctou_asi_")
    try:
        out = tmp / "o.xlsx"
        wb = Workbook()
        committed = artifact_store.atomic_save_if(wb, out, lambda: False)
        check("proceed() False -> not committed, destination NOT written",
              committed is False and not out.exists())
        wb2 = Workbook()
        committed2 = artifact_store.atomic_save_if(wb2, out, lambda: True)
        check("proceed() True -> committed, destination written",
              committed2 is True and out.is_file())
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ----------------------------------------------------------------------------- #
# commit_workbook (the comparator write path)
# ----------------------------------------------------------------------------- #
def commit_workbook_checks():
    print("commit_workbook (comparator path):")

    def run_case(name, *, pre_exist, appears, confirm_returns, expect_status,
                 expect_confirm_calls):
        tmp = _tmp("tsmis_toctou_cw_")
        try:
            final = tmp / "out.xlsx"
            if pre_exist:
                final.write_text("PRIOR", encoding="utf-8")
            calls = []

            def confirm(_p):
                calls.append(_p)
                return confirm_returns

            def produce(temp_path):
                _write_valid_xlsx(temp_path)
                if appears:
                    final.write_text("APPEARED", encoding="utf-8")
                return ConsolidateResult(status="ok", output_path=str(temp_path))

            res = artifact_store.commit_workbook(final, produce, confirm_overwrite=confirm)
            check(f"{name}: status == {expect_status}", res.status == expect_status)
            check(f"{name}: confirm called {expect_confirm_calls}x",
                  len(calls) == expect_confirm_calls)
            if expect_status == "cancelled" and appears:
                check(f"{name}: the APPEARED file was preserved",
                      final.read_text(encoding="utf-8") == "APPEARED")
            if expect_status == "ok":
                check(f"{name}: a valid workbook was committed",
                      artifact_store._openable_xlsx(final))
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    run_case("appeared+decline", pre_exist=False, appears=True, confirm_returns=False,
             expect_status="cancelled", expect_confirm_calls=1)
    run_case("appeared+accept", pre_exist=False, appears=True, confirm_returns=True,
             expect_status="ok", expect_confirm_calls=1)
    run_case("no-appearance", pre_exist=False, appears=False, confirm_returns=False,
             expect_status="ok", expect_confirm_calls=0)
    run_case("pre-existing-accept", pre_exist=True, appears=False, confirm_returns=True,
             expect_status="ok", expect_confirm_calls=1)


# ----------------------------------------------------------------------------- #
# consolidate_xlsx — real gate via the decorate_workbook hook (appears mid-build)
# ----------------------------------------------------------------------------- #
def consolidate_xlsx_checks():
    print("consolidate_xlsx (plain XLSX consolidator, real gate):")
    tmp = _tmp("tsmis_toctou_xlsx_")
    try:
        in_dir = tmp / "in"
        in_dir.mkdir()
        _write_valid_xlsx(in_dir / "route1.xlsx", sheet="Data",
                          header=("PM", "Val"), rows=(("000.001", "x"),))
        out_path = tmp / "combined.xlsx"
        calls = []

        def confirm(_p):
            calls.append(_p)
            return False

        def appear(_wb):
            out_path.write_text("APPEARED", encoding="utf-8")

        res = consolidate_xlsx_base.consolidate_xlsx(
            input_dir=in_dir, out_path=out_path, sheet_name="Data",
            report_name="Test", title="Test", events=Events(),
            confirm_overwrite=confirm, decorate_workbook=appear)
        check("appeared-then-decline -> cancelled", res.status == "cancelled")
        check("confirm called exactly once", len(calls) == 1)
        check("the APPEARED file was preserved", out_path.read_text(encoding="utf-8") == "APPEARED")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ----------------------------------------------------------------------------- #
# the four direct/converter consolidators — appearance AT the final save (P12-B01)
# ----------------------------------------------------------------------------- #
def _assert_late_save(label, calls, res, out_path):
    check(f"{label}: appeared-at-final-save + decline -> cancelled",
          res.status == "cancelled")
    check(f"{label}: confirm called once (the gate fired at the save)", len(calls) == 1)
    check(f"{label}: the APPEARED file was preserved (not overwritten)",
          out_path.exists() and out_path.read_text(encoding="utf-8") == "APPEARED")


def ramp_summary_late_save():
    print("consolidate_ramp_summary (gate at the final save):")
    M = consolidate_ramp_summary
    tmp = _tmp("tsmis_toctou_ramp_")
    try:
        in_dir = tmp / "in"
        in_dir.mkdir()
        (in_dir / "r1.pdf").write_bytes(b"%PDF-1.4")
        out_path = tmp / "ramp.xlsx"
        calls = []
        saved = (M.parse_pdf, M.record_has_data)
        M.parse_pdf = lambda _p: {"route": "001"}
        M.record_has_data = lambda _rec: True
        try:
            res = _run_with_late_save_appearance(out_path, lambda: M.consolidate(
                events=Events(), confirm_overwrite=lambda _p: calls.append(_p) or False,
                input_dir=in_dir, out_path=out_path))
        finally:
            M.parse_pdf, M.record_has_data = saved
        _assert_late_save("ramp_summary", calls, res, out_path)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def intersection_summary_late_save():
    print("consolidate_intersection_summary (gate at the final save):")
    M = consolidate_intersection_summary
    tmp = _tmp("tsmis_toctou_int_")
    try:
        in_dir = tmp / "in"
        in_dir.mkdir()
        _write_valid_xlsx(in_dir / "i1.xlsx")
        out_path = tmp / "int.xlsx"
        calls = []
        saved = (M.parse_route, M.record_has_data)
        M.parse_route = lambda _p: ("001", {}, 0)
        M.record_has_data = lambda _rec: True
        try:
            res = _run_with_late_save_appearance(out_path, lambda: M.consolidate(
                events=Events(), confirm_overwrite=lambda _p: calls.append(_p) or False,
                input_dir=in_dir, out_path=out_path))
        finally:
            M.parse_route, M.record_has_data = saved
        _assert_late_save("intersection_summary", calls, res, out_path)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def tsn_highway_sequence_late_save():
    print("consolidate_tsn_highway_sequence (gate at the final save):")
    M = consolidate_tsn_highway_sequence
    tmp = _tmp("tsmis_toctou_tsnseq_")
    try:
        in_dir = tmp / "in"
        in_dir.mkdir()
        (in_dir / "d1.pdf").write_bytes(b"%PDF-1.4")
        out_path = tmp / "seq.xlsx"
        calls = []
        saved = M.parse_pdf
        saved_universe = M.tdc.require_exact_universe
        _row = {"county": "04", "pm": "0.000", "city": "", "hg": "", "ft": "",
                "dist": "", "description": ""}
        M.parse_pdf = lambda _p, _e, pdf_name="": ("01", {"001": [_row]})
        M.tdc.require_exact_universe = lambda claimed: tuple(claimed)
        try:
            res = _run_with_late_save_appearance(out_path, lambda: M.consolidate(
                events=Events(), confirm_overwrite=lambda _p: calls.append(_p) or False,
                input_dir=in_dir, out_path=out_path))
        finally:
            M.parse_pdf = saved
            M.tdc.require_exact_universe = saved_universe
        _assert_late_save("tsn_highway_sequence", calls, res, out_path)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def converter_late_save():
    print("consolidate_tsmis_highway_log_pdf (per-route converter, gate at the combine):")
    M = consolidate_tsmis_highway_log_pdf
    tmp = _tmp("tsmis_toctou_conv_")
    try:
        in_dir = tmp / "in"
        in_dir.mkdir()
        (in_dir / "route001.pdf").write_bytes(b"%PDF-1.4")
        out_path = tmp / "combined.xlsx"
        conv = tmp / "conv"
        calls = []
        saved = M.parse_pdf
        M.parse_pdf = lambda _p, _e, pdf_name="": (
            "001", [["000.001"] + ["x"] * 30],
            {"emitted": 1, "pages": 1, "skipped_no_geometry": 0,
             "stale_geometry_pages": 0, "carried_validated_pages": 0})
        try:
            res = _run_with_late_save_appearance(out_path, lambda: M.consolidate(
                events=Events(), confirm_overwrite=lambda _p: calls.append(_p) or False,
                input_dir=in_dir, out_path=out_path, converted_dir=conv))
        finally:
            M.parse_pdf = saved
        _assert_late_save("converter", calls, res, out_path)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main():
    print("Consolidate-overwrite TOCTOU re-check:")
    unit_checks()
    atomic_save_if_checks()
    commit_workbook_checks()
    consolidate_xlsx_checks()
    ramp_summary_late_save()
    intersection_summary_late_save()
    tsn_highway_sequence_late_save()
    converter_late_save()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL CONSOLIDATE-TOCTOU CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
