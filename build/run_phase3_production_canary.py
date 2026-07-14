"""Run the bound CORE-ID-78 production comparator and Excel parity canary.

Unlike the independent oracle, this runner deliberately imports the current
production consolidator, normalizer, comparator, contract, and persistence
layers.  Its job is to prove that those layers reproduce the already-bound
oracle exactly, persist the duplicate-pairing trace, and emit a formulas
workbook whose installed-Excel results equal the values twin.

The evidence directory must not already exist.  Raw corpus files are read only;
every derived artifact is written under the supplied evidence directory.
"""
from __future__ import annotations

import argparse
from dataclasses import asdict
import hashlib
from itertools import zip_longest
import json
from pathlib import Path
import shutil
import sys
import time

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT / "build")]

import compare_intersection_detail_tsn as comparator  # noqa: E402
import consolidation_meta  # noqa: E402
import consolidate_intersection_detail as consolidator  # noqa: E402
import outcome  # noqa: E402
import phase3_intersection_detail_oracle as binding  # noqa: E402
import tsn_load_intersection_detail as normalizer  # noqa: E402
from comparison_contract import PairingTrace  # noqa: E402
from events import Events  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


CANARY_ID = "CORE-ID-78-XLSX-TSN"
EXPECTED_MANIFEST = (
    "9d1c0ae4f9bc8de098497695cd87d3c543dba01e34cb9f4b03cb883791b52bd6"
)
EXPECTED_MEMBERS = 218
EXPECTED_SOURCE_BYTES = 26_384_760


class CanaryEvents(Events):
    def __init__(self):
        super().__init__()
        self.lines = []

    def on_log(self, line=""):
        text = str(line)
        self.lines.append(text)
        if text and ("Writing" in text or "Building" in text or "rows:" in text):
            print(text, flush=True)


def _sha256(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _artifact(path):
    path = Path(path)
    facts = consolidation_meta._bound_file_digest(path)
    if facts is None:
        raise AssertionError(f"artifact changed or became unreadable: {path.name}")
    return {"bytes": facts["size"], "sha256": facts["sha256"]}


def _artifact_table(paths):
    """Identity-bind every artifact and reject filename-key collisions."""
    result = {}
    names = set()
    for value in paths:
        path = Path(value)
        name_key = path.name.casefold()
        if name_key in names:
            raise AssertionError(
                f"case-insensitive evidence artifact collision: {path.name!r}")
        names.add(name_key)
        result[path.name] = _artifact(path)
    return result


def _canonical_json_bytes(value):
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
        allow_nan=False).encode("utf-8")


def _comparison_payload_evidence(formulas, values, comparison_result):
    """Independently bind schema-v3 envelopes, payload, and exact chunk set."""
    workbooks = (Path(formulas), Path(values))
    if workbooks[0].parent != workbooks[1].parent:
        raise AssertionError("comparison twins do not share one evidence directory")

    typed = comparison_result.comparison_outcome
    generation = comparison_result.artifact_generation
    manifests = []
    envelopes = []
    expected_artifacts = {}
    for workbook in workbooks:
        sidecar = consolidation_meta.meta_path(workbook)
        sentinel = sidecar.with_name(sidecar.name + ".tmp")
        try:
            sentinel.lstat()
        except (FileNotFoundError, NotADirectoryError):
            pass
        else:
            raise AssertionError(
                f"comparison publication sentinel remains: {sentinel.name}")

        sidecar_raw = consolidation_meta._read_bound_bytes(
            sidecar, consolidation_meta._MAX_COMPARISON_SIDECAR_BYTES,
            "production canary comparison sidecar")
        if not isinstance(sidecar_raw, bytes):
            raise AssertionError(f"comparison sidecar is absent: {sidecar.name}")
        envelope = consolidation_meta._decode_strict_json(
            sidecar_raw, "production canary comparison sidecar")
        if (not isinstance(envelope, dict)
                or envelope.get("comparison_schema_version") != 3):
            raise AssertionError(
                f"production canary did not exercise schema v3: {sidecar.name}")
        manifest = consolidation_meta._strict_payload_manifest(
            envelope.get("comparison_payload"))
        envelopes.append(envelope)
        manifests.append(manifest)
        expected_artifacts[sidecar.name] = {
            "bytes": len(sidecar_raw),
            "sha256": hashlib.sha256(sidecar_raw).hexdigest(),
        }

    if manifests[0] != manifests[1]:
        raise AssertionError("comparison twins reference different payload manifests")
    for key in ("completion", "skipped_inputs", "failed_inputs",
                "artifact_generation"):
        if envelopes[0].get(key) != envelopes[1].get(key):
            raise AssertionError(
                f"comparison twins disagree on publication field {key!r}")

    manifest = manifests[0]
    outcome_raw = _canonical_json_bytes(typed.to_dict())
    outcome_sha = hashlib.sha256(outcome_raw).hexdigest()
    if (manifest["decoded_size"] != len(outcome_raw)
            or manifest["decoded_sha256"] != outcome_sha):
        raise AssertionError(
            "persisted payload manifest does not bind the returned typed outcome")

    completion = envelopes[0]["completion"]
    skipped = envelopes[0]["skipped_inputs"]
    failed = envelopes[0]["failed_inputs"]
    generation_payload = generation.to_dict()
    binding_raw = _canonical_json_bytes({
        "decoded_sha256": outcome_sha,
        "completion": completion,
        "skipped_inputs": skipped,
        "failed_inputs": failed,
        "artifact_generation": generation_payload,
    })
    binding_sha = hashlib.sha256(binding_raw).hexdigest()
    if manifest["binding_sha256"] != binding_sha:
        raise AssertionError(
            "payload binding does not match the returned artifact generation")

    parent = workbooks[0].parent
    referenced = {item["relative_path"] for item in manifest["chunks"]}
    present = {
        entry.name for entry in parent.iterdir()
        if consolidation_meta._PAYLOAD_BASENAME_RE.fullmatch(entry.name)
    }
    if present != referenced:
        raise AssertionError(
            "payload artifact set is not exact: "
            f"present={sorted(present)!r}, referenced={sorted(referenced)!r}")

    chunk_paths = []
    for descriptor in manifest["chunks"]:
        chunk = parent / descriptor["relative_path"]
        raw = consolidation_meta._read_bound_bytes(
            chunk, consolidation_meta._MAX_COMPARISON_PAYLOAD_CHUNK_BYTES,
            "production canary payload chunk")
        if (not isinstance(raw, bytes)
                or len(raw) != descriptor["size"]
                or hashlib.sha256(raw).hexdigest() != descriptor["sha256"]):
            raise AssertionError(
                f"payload chunk does not match its descriptor: {chunk.name}")
        chunk_paths.append(chunk)
        expected_artifacts[chunk.name] = {
            "bytes": len(raw),
            "sha256": hashlib.sha256(raw).hexdigest(),
        }

    # Perform the complete production reader/result agreement check after the
    # exact envelope/chunk bytes above were captured. The caller then proves its
    # one final artifact table still describes those same bytes.
    for workbook in workbooks:
        consolidation_meta.require_published_comparison(
            workbook, comparison_result)

    lock_path = parent / consolidation_meta._COMPARISON_PUBLICATION_LOCK_NAME
    lock_facts = _artifact(lock_path)
    expected_artifacts[lock_path.name] = lock_facts

    return {
        "record_schema_version": envelopes[0]["schema_version"],
        "comparison_schema_version": 3,
        "completion": completion,
        "skipped_inputs": skipped,
        "failed_inputs": failed,
        "artifact_generation": generation_payload,
        "comparison_payload": manifest,
        "publication_lock": {
            "relative_path": lock_path.name,
            **lock_facts,
        },
    }, tuple(chunk_paths), lock_path, expected_artifacts


def _write_json(path, value):
    Path(path).write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8", newline="\n")


def _write_jsonl(path, records):
    with Path(path).open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(
                record, ensure_ascii=False, sort_keys=True,
                separators=(",", ":")))
            handle.write("\n")


def _manifest_assert(snapshot, label):
    actual = (len(snapshot.records), snapshot.source_bytes, snapshot.sha256)
    expected = (EXPECTED_MEMBERS, EXPECTED_SOURCE_BYTES, EXPECTED_MANIFEST)
    if actual != expected:
        raise AssertionError(f"{label} raw manifest changed: {actual!r}")


def _expected_counts(oracle_result):
    value = json.loads(Path(oracle_result).read_text(encoding="utf-8"))
    if (value.get("canary_id") != CANARY_ID
            or value.get("completion") != "complete"
            or value.get("input_binding", {}).get("manifest_sha256")
            != EXPECTED_MANIFEST):
        raise AssertionError("oracle result is not the bound complete ID78 result")
    return value


def _production_counts(typed):
    counts = typed.counts
    per_field = {}
    for token, value in counts.per_field_counts.items():
        _index, separator, label = token.partition(":")
        if not separator or not label:
            raise AssertionError(f"malformed production per-field token: {token!r}")
        if value:
            if label in per_field:
                raise AssertionError(f"duplicate production field label: {label!r}")
            per_field[label] = value
    return {
        "known": counts.known,
        "paired_rows": counts.paired_rows,
        "side_a_only_rows": counts.side_a_only_rows,
        "side_b_only_rows": counts.side_b_only_rows,
        "differing_rows": counts.differing_rows,
        "differing_cells": counts.differing_cells,
        "per_field_counts": per_field,
        "asserted_cells": counts.asserted_cells,
        "context_cells": counts.context_cells,
    }


def _trace_record(trace):
    if not isinstance(trace, PairingTrace):
        raise TypeError("production trace is not typed")
    return {
        "algorithm": trace.algorithm,
        "assignment_vector": list(trace.assignment_vector),
        "exact": trace.exact,
        "key": list(trace.key_components),
        "matrix_cells": trace.matrix_cells,
        "pairs": [[pair.side_a_index, pair.side_b_index, pair.cost]
                  for pair in trace.pairs],
        "positional_cost": trace.positional_cost,
        "quality": trace.quality,
        "side_a_indices": list(trace.side_a_indices),
        "side_a_size": trace.side_a_size,
        "side_b_indices": list(trace.side_b_indices),
        "side_b_size": trace.side_b_size,
        "smaller_side": trace.smaller_side,
        "total_cost": trace.total_cost,
    }


def _oracle_duplicate_traces(path):
    records = {}
    with Path(path).open("r", encoding="utf-8") as handle:
        for line in handle:
            record = json.loads(line)
            if record["side_a_size"] == record["side_b_size"] == 1:
                continue
            key = tuple(value["text"] for value in record["key"])
            records[key] = record
    return records


def _assert_trace_parity(production, oracle_trace_path):
    expected = _oracle_duplicate_traces(oracle_trace_path)
    actual = {tuple(trace.key_components): trace for trace in production}
    if len(actual) != len(production):
        raise AssertionError("production emitted duplicate trace keys")
    if set(actual) != set(expected):
        raise AssertionError(
            f"duplicate trace key drift: production={len(actual)}, "
            f"oracle={len(expected)}")
    mismatches = []
    for key in sorted(expected):
        left, right = actual[key], expected[key]
        checks = {
            "sizes": ((left.side_a_size, left.side_b_size),
                      (right["side_a_size"], right["side_b_size"])),
            "smaller": (left.smaller_side, right["smaller_side"]),
            "vector": (list(left.assignment_vector), right["assignment_vector"]),
            "pairs": ([[pair.side_a_index, pair.side_b_index]
                       for pair in left.pairs], right["source_pairs"]),
            "cost": (left.total_cost, right["total_cost"]),
            "algorithm": (left.algorithm, right["algorithm"]),
            "quality": (left.quality, right["quality"]),
            "exact": (left.exact, right["exact"]),
        }
        bad = {name: values for name, values in checks.items()
               if values[0] != values[1]}
        if bad:
            mismatches.append((key, bad))
    if mismatches:
        raise AssertionError(f"production/oracle pairing mismatch: {mismatches[:3]!r}")


def _excel_recalculate(paths):
    try:
        import win32com.client as win32
    except Exception as exc:
        raise RuntimeError("installed Excel COM dependency is unavailable") from exc
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        for path in paths:
            print(f"Excel full rebuild: {Path(path).name}", flush=True)
            book = None
            try:
                book = excel.Workbooks.Open(
                    str(Path(path).resolve()), UpdateLinks=0, ReadOnly=False)
                excel.CalculateFullRebuild()
                book.Save()
                book.Close(SaveChanges=False)
                book = None
            finally:
                if book is not None:
                    book.Close(SaveChanges=False)
    finally:
        if excel is not None:
            excel.Quit()


def _header_index(header, name):
    positions = [index for index, value in enumerate(header) if value == name]
    if len(positions) != 1:
        raise AssertionError(f"expected one {name!r} header, found {positions!r}")
    return positions[0]


def _comparison_excel_parity(formulas_path, values_path, typed):
    formulas = load_workbook(formulas_path, read_only=True, data_only=True)
    values = load_workbook(values_path, read_only=True, data_only=True)
    try:
        left = formulas["Comparison"].iter_rows(values_only=True)
        right = values["Comparison"].iter_rows(values_only=True)
        left_header, right_header = tuple(next(left)), tuple(next(right))
        if left_header != right_header:
            raise AssertionError("Comparison headers differ between twins")
        status_col = _header_index(left_header, "Status")
        diffs_col = _header_index(left_header, "Diffs")
        state_cols = [index for index, value in enumerate(left_header)
                      if isinstance(value, str)
                      and value.startswith("__CMP_E1_STATE_V1_")]
        if not state_cols:
            raise AssertionError("Comparison has no typed state-mask columns")
        rows = paired = a_only = b_only = diff_rows = diff_cells = 0
        mismatch = []
        for row_no, (a_row, b_row) in enumerate(
                zip_longest(left, right), start=2):
            if a_row is None or b_row is None:
                mismatch.append((row_no, "row-count", a_row, b_row))
                break
            a_row, b_row = tuple(a_row), tuple(b_row)
            if a_row != b_row:
                locations = [index for index, pair in enumerate(zip(a_row, b_row))
                             if pair[0] != pair[1]]
                mismatch.append((row_no, locations[:8]))
                if len(mismatch) >= 8:
                    break
            if not a_row or all(value in (None, "") for value in a_row):
                continue
            rows += 1
            status = a_row[status_col]
            mask = "".join(str(a_row[index] or "") for index in state_cols)
            if status == "Both":
                paired += 1
                observed = int(a_row[diffs_col] or 0)
                if observed != mask.count("D") or set(mask) - set("EDN"):
                    raise AssertionError(
                        f"formula row {row_no} has inconsistent typed state {mask!r}")
                diff_cells += observed
                diff_rows += observed > 0
            else:
                if set(mask) != {"U"} or a_row[diffs_col] not in (None, ""):
                    raise AssertionError(
                        f"one-sided formula row {row_no} has invalid state {mask!r}")
                if status == "TSMIS only":
                    a_only += 1
                elif status == "TSN only":
                    b_only += 1
                else:
                    raise AssertionError(f"unknown Comparison status {status!r}")
        if mismatch:
            raise AssertionError(f"formula/values Comparison parity failed: {mismatch!r}")
        observed = (paired, a_only, b_only, diff_rows, diff_cells)
        expected = (typed.counts.paired_rows, typed.counts.side_a_only_rows,
                    typed.counts.side_b_only_rows, typed.counts.differing_rows,
                    typed.counts.differing_cells)
        if observed != expected or rows != typed.counts.union_rows:
            raise AssertionError(
                f"calculated Comparison counts {observed!r}/{rows} != {expected!r}")
        return {"rows": rows, "state_chunks": len(state_cols)}
    finally:
        formulas.close()
        values.close()


def _summary_map(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Summary"]
        result = {}
        for row in ws.iter_rows(values_only=True):
            label = row[1] if len(row) > 1 else None
            value = row[2] if len(row) > 2 else None
            if isinstance(label, str) and value is not None:
                result.setdefault(label, []).append(value)
        return result
    finally:
        wb.close()


def _self_check_labels(path):
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["Summary"]
        labels = []
        for row in ws.iter_rows():
            label = row[1].value if len(row) > 1 else None
            formula = row[2].value if len(row) > 2 else None
            if (isinstance(label, str) and isinstance(formula, str)
                    and formula.startswith("=IF(")
                    and '"OK","CHECK"' in formula):
                labels.append(label)
        return tuple(labels)
    finally:
        wb.close()


def _summary_excel_parity(formulas_path, values_path):
    left, right = _summary_map(formulas_path), _summary_map(values_path)
    common = set(left) & set(right)
    mismatches = {label: (left[label], right[label]) for label in common
                  if left[label] != right[label]}
    if mismatches:
        raise AssertionError(f"Summary twin mismatch: {list(mismatches.items())[:8]!r}")
    labels = _self_check_labels(formulas_path)
    if len(labels) < 9:
        raise AssertionError(f"too few Summary self-checks: {labels!r}")
    failures = [(label, left.get(label), right.get(label)) for label in labels
                if left.get(label) != ["OK"] or right.get(label) != ["OK"]]
    if failures:
        raise AssertionError(f"Summary self-check failure: {failures!r}")
    return {"shared_numeric_labels": len(common), "self_checks": len(labels)}


def _spot_check(path, expected_fields):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Spot Check"]
        header_row = header_col = None
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_index, value in enumerate(row, start=1):
                if value == "Agree?":
                    header_row, header_col = row_index, col_index
        if header_row is None:
            raise AssertionError("Spot Check has no Agree? column")
        verdicts = [ws.cell(row, header_col).value
                    for row in range(header_row + 1, header_row + 1 + expected_fields)]
        if verdicts != ["OK"] * expected_fields:
            raise AssertionError(f"Spot Check parity failed: {verdicts!r}")
        return len(verdicts)
    finally:
        wb.close()


def _report_view_diffs(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Report View"]
        return sum(value for (value,) in ws.iter_rows(
            min_col=2, max_col=2, min_row=5, values_only=True)
            if isinstance(value, (int, float)) and not isinstance(value, bool))
    finally:
        wb.close()


def _helper_parity(formulas_path, values_path):
    formulas = load_workbook(formulas_path, read_only=True, data_only=True)
    values = load_workbook(values_path, read_only=True, data_only=True)
    try:
        compared = 0
        for sheet in ("TSMIS", "TSN"):
            left = formulas[sheet].iter_rows(values_only=True)
            right = values[sheet].iter_rows(values_only=True)
            lh, rh = tuple(next(left)), tuple(next(right))
            if lh != rh:
                raise AssertionError(f"{sheet} twin headers differ")
            columns = [index for index, label in enumerate(lh)
                       if label == "Key (helper)"
                       or (isinstance(label, str)
                           and label.startswith("__CMP_E1_MW_V1_"))]
            if not columns:
                raise AssertionError(f"{sheet} has no helper columns")
            for row_no, (a_row, b_row) in enumerate(
                    zip_longest(left, right), start=2):
                if a_row is None or b_row is None:
                    raise AssertionError(f"{sheet} helper row count differs")
                for column in columns:
                    if a_row[column] != b_row[column]:
                        raise AssertionError(
                            f"{sheet} helper mismatch row {row_no}, col {column + 1}")
                    compared += 1
        return compared
    finally:
        formulas.close()
        values.close()


def _code_manifest():
    paths = (
        "scripts/compare_core.py",
        "scripts/comparison_contract.py",
        "scripts/compare_intersection_detail_tsn.py",
        "scripts/compare_tsn_common.py",
        "scripts/consolidate_intersection_detail.py",
        "scripts/consolidate_xlsx_base.py",
        "scripts/tsn_load_intersection_detail.py",
        "scripts/tsn_library.py",
        "scripts/consolidation_meta.py",
        "scripts/artifact_store.py",
        "build/run_phase3_production_canary.py",
    )
    return {path: _sha256(ROOT / path) for path in paths}


def run(args):
    started = time.perf_counter()
    evidence = Path(args.evidence_dir).resolve()
    evidence.mkdir(parents=True, exist_ok=False)
    oracle_result_path = Path(args.oracle_result).resolve()
    oracle_trace_path = oracle_result_path.with_name(
        f"{CANARY_ID}-pairing-trace.jsonl")
    oracle = _expected_counts(oracle_result_path)

    selection = binding.select_corpus(args.corpus_root)
    pre = binding.capture_manifest(selection)
    _manifest_assert(pre, "pre-run")
    raw_manifest_path = evidence / f"{CANARY_ID}-raw-manifest-v1.tsv"
    raw_manifest_path.write_bytes(pre.serialized)

    events = CanaryEvents()
    consolidated = evidence / "tsmis-intersection-detail-consolidated.xlsx"
    print("Building production TSMIS consolidation…", flush=True)
    consolidated_result = consolidator.consolidate(
        events=events, confirm_overwrite=lambda _path: True,
        input_dir=selection.id78_root / "intersection_detail",
        out_path=consolidated)
    if (consolidated_result.status != "ok"
            or consolidated_result.completion != outcome.COMPLETE
            or consolidated_result.skipped_inputs
            or consolidated_result.failed_inputs):
        raise AssertionError(f"TSMIS consolidation incomplete: {consolidated_result!r}")
    if not consolidation_meta.write_outcome(
            consolidated, consolidated_result,
            extra={"production_canary": CANARY_ID}):
        raise AssertionError("could not persist TSMIS consolidation outcome")

    normalized = evidence / "tsn-intersection-detail-normalized.xlsx"
    print("Building production TSN normalized library…", flush=True)
    normalized_result = normalizer.build_into(
        selection.tsn_file.parent, normalized, events=events,
        confirm_overwrite=lambda _path: True)
    if normalized_result.status != "ok":
        raise AssertionError(f"TSN normalization failed: {normalized_result!r}")
    if not consolidation_meta.write_outcome(
            normalized, normalized_result,
            extra={"production_canary": CANARY_ID,
                   "tsn_normalization_version": "v3"}):
        raise AssertionError("could not persist TSN normalization outcome")
    for source in (consolidated, normalized):
        record = consolidation_meta.read_outcome(source)
        if (record is None or not record.trusted
                or record.completion != outcome.COMPLETE):
            raise AssertionError(f"derived input outcome is not trusted: {source.name}")

    before_compare = binding.capture_manifest(selection)
    if before_compare.serialized != pre.serialized:
        raise AssertionError("raw corpus changed before production comparison")

    formulas = evidence / f"{CANARY_ID}-formulas.xlsx"
    print("Building production formulas + values comparison twins…", flush=True)
    comparison_result = comparator.compare(
        consolidated, normalized, formulas, events=events,
        confirm_overwrite=lambda _path: True, mode="both")
    if comparison_result.status != "ok":
        raise AssertionError(f"production comparison failed: {comparison_result!r}")
    typed = comparison_result.comparison_outcome
    if (typed.status != "ok" or typed.completion != "complete"
            or typed.verdict != "diff" or typed.pairing_quality != "exact"
            or typed.capped_group_diagnostics or not typed.is_complete):
        raise AssertionError(f"production typed outcome is not exact/complete: {typed!r}")

    production_counts = _production_counts(typed)
    if production_counts != oracle["counts"]:
        raise AssertionError(
            f"production/oracle count mismatch: {production_counts!r} != "
            f"{oracle['counts']!r}")
    if len(typed.pairing_trace) != oracle["pairing"]["duplicate_groups"]:
        raise AssertionError(
            f"production duplicate trace count {len(typed.pairing_trace)} != "
            f"oracle {oracle['pairing']['duplicate_groups']}")
    _assert_trace_parity(typed.pairing_trace, oracle_trace_path)
    trace_path = evidence / f"{CANARY_ID}-production-pairing-trace.jsonl"
    _write_jsonl(trace_path, (_trace_record(trace) for trace in typed.pairing_trace))

    values = formulas.with_name(f"{formulas.stem} (values){formulas.suffix}")
    if not formulas.is_file() or not values.is_file():
        raise AssertionError("comparison twins were not both published")
    for workbook in (formulas, values):
        try:
            consolidation_meta.require_published_comparison(
                workbook, comparison_result)
        except ValueError as error:
            raise AssertionError(
                f"comparison sidecar is not exact/trusted: {workbook.name}"
            ) from error

    formulas_excel = evidence / f"{CANARY_ID}-formulas-excel-recalculated.xlsx"
    values_excel = evidence / f"{CANARY_ID}-values-excel-recalculated.xlsx"
    shutil.copy2(formulas, formulas_excel)
    shutil.copy2(values, values_excel)
    _excel_recalculate((formulas_excel, values_excel))

    parity = {
        "comparison": _comparison_excel_parity(formulas_excel, values_excel, typed),
        "summary": _summary_excel_parity(formulas_excel, values_excel),
        "spot_check_fields": _spot_check(formulas_excel, len(comparator._SCHEMA.field_indices)),
        "report_view_diffs_formulas": _report_view_diffs(formulas_excel),
        "report_view_diffs_values": _report_view_diffs(values_excel),
        "helper_cells_compared": _helper_parity(formulas_excel, values_excel),
    }
    expected_report_view = 2 * typed.counts.differing_cells
    if (parity["report_view_diffs_formulas"] != expected_report_view
            or parity["report_view_diffs_values"] != expected_report_view):
        raise AssertionError(
            f"Report View Diffs do not equal 2× typed truth: {parity!r}")

    post = binding.capture_manifest(selection)
    if post.serialized != pre.serialized:
        raise AssertionError("raw corpus changed during production canary")

    events_path = evidence / f"{CANARY_ID}-production-events.log"
    events_path.write_text("\n".join(events.lines) + "\n", encoding="utf-8")

    # Recalculation occurred on evidence copies. At the final evidence boundary,
    # rebind the committed originals, strict envelopes, returned result, and
    # every shared payload chunk before hashing the complete artifact set once.
    (comparison_publication, payload_paths, publication_lock_path,
     expected_publication_artifacts) = _comparison_payload_evidence(
         formulas, values, comparison_result)
    artifact_paths = (
        raw_manifest_path,
        consolidated, consolidation_meta.meta_path(consolidated),
        normalized, consolidation_meta.meta_path(normalized),
        formulas, consolidation_meta.meta_path(formulas),
        values, consolidation_meta.meta_path(values),
        formulas_excel, values_excel, trace_path, events_path,
    ) + payload_paths + (publication_lock_path,)
    artifacts = _artifact_table(artifact_paths)
    for name, expected in expected_publication_artifacts.items():
        if artifacts.get(name) != expected:
            raise AssertionError(
                f"final artifact facts changed after publication binding: {name}")
    for member in comparison_result.artifact_generation.members:
        expected = {"bytes": member["size"], "sha256": member["sha256"]}
        if artifacts.get(member["relative_path"]) != expected:
            raise AssertionError(
                "final workbook artifact disagrees with ArtifactGeneration: "
                f"{member['relative_path']}")
    if artifacts[raw_manifest_path.name]["sha256"] != pre.sha256:
        raise AssertionError("raw-manifest evidence bytes disagree with input binding")
    result = {
        "canary_id": CANARY_ID,
        "completion": "complete",
        "verdict": typed.verdict,
        "counts": production_counts,
        "pairing": {
            "quality": typed.pairing_quality,
            "duplicate_trace_records": len(typed.pairing_trace),
            "capped_groups": len(typed.capped_group_diagnostics),
            "trace_sha256": artifacts[trace_path.name]["sha256"],
        },
        "input_binding": {
            "members": len(pre.records),
            "source_bytes": pre.source_bytes,
            "pre_manifest_sha256": pre.sha256,
            "before_compare_manifest_sha256": before_compare.sha256,
            "post_manifest_sha256": post.sha256,
        },
        "excel_parity": parity,
        "comparison_publication": comparison_publication,
        "artifacts": artifacts,
        "production_source_sha256": _code_manifest(),
        "oracle_result_sha256": _sha256(oracle_result_path),
        "oracle_trace_sha256": _sha256(oracle_trace_path),
        "elapsed_seconds": round(time.perf_counter() - started, 3),
    }
    result_path = evidence / f"{CANARY_ID}-production-result.json"
    _write_json(result_path, result)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True), flush=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--corpus-root", required=True)
    parser.add_argument("--evidence-dir", required=True)
    parser.add_argument("--oracle-result", required=True)
    args = parser.parse_args()
    run(args)


if __name__ == "__main__":
    main()
