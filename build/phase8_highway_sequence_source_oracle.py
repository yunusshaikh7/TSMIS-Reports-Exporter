#!/usr/bin/env python3
"""Independent Highway Sequence TSMIS Excel/PDF source oracle.

The oracle reads only the immutable capture produced by
``phase8_highway_sequence_capture.py``.  It does not import the application's
Highway Sequence parser, consolidator, comparison adapter, evidence adapter, or
schema.  Excel is read positionally with strict workbook/sheet/row contracts.
PDF rows are reconstructed from pdfplumber words against a statewide fixed-grid
model; production instead derives windows from each page's header.

The first job of this module is source parity, not product certification.  It
preserves every current Excel/PDF difference, including format-specific EQUATES
representations, before either side is projected into comparison semantics.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor
from dataclasses import asdict, dataclass
import hashlib
import json
import logging
import math
import os
from pathlib import Path
import re
from typing import Iterable, Mapping, Sequence

import pdfplumber
from openpyxl import load_workbook


logging.getLogger("pdfminer").setLevel(logging.ERROR)

PRIVATE_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd"
    r"\phase8_highway_sequence_private_sources_r1"
)
CAPTURE_MANIFEST = PRIVATE_ROOT / "capture_manifest.json"
DEFAULT_OUTPUT = PRIVATE_ROOT.parent / "phase8_highway_sequence_source_oracle_draft_r1.json"
DEFAULT_ROW_CACHE = PRIVATE_ROOT.parent / "phase8_highway_sequence_source_rows_draft_r1.json"

CAPTURE_MANIFEST_BINDING = {
    "bytes": 145_434,
    "sha256": "6f41566c350797f135916e0d5b9f0de434e000faa5882ae1309d866f87cc6534",
    "stable_capture_sha256": "e6fd69838aef7fca34fd1c5cdd8b79e5ddb1c72e229bcf1928f63934ac91eceb",
}
TREE_BINDINGS = {
    "current_tsmis_excel": (252, 24_634_973, "31a13ebc388951fdcadbba69d9188218af4548dd56d68c91e09f96bcb41765c8", ".xlsx"),
    "current_tsmis_pdf": (252, 39_236_260, "072e538e5ebcbf015ec719565f003fb72027973a11d63c42f123802d8856dfa7", ".pdf"),
    "historical_tsmis_excel_7_8": (252, 24_634_499, "4bb040280bab17fd14283aa20178d189b4e499291eea1345adba0e0bb7f72c4f", ".xlsx"),
    "historical_tsmis_pdf_7_9": (252, 39_236_260, "072e538e5ebcbf015ec719565f003fb72027973a11d63c42f123802d8856dfa7", ".pdf"),
    "authoritative_tsn_pdf": (12, 3_866_949, "91d63fc20e82c8368044a9ef00224cd4b9b55309af55109fd34e4dacba7e72a2", ".pdf"),
}

HEADERS = (
    "County", "City", None, "PM", None, "HG", "FT",
    "Distance To Next Point", "Description",
)
FIELD_NAMES = (
    "County", "City", "PM Prefix", "PM", "PM Suffix", "HG", "FT",
    "Distance To Next Point", "Description",
)
SHEET_NAME = "Highway Locations"
MEMBER_RE = re.compile(r"^highway_sequence_route_(\d{3}[A-Za-z]?)\.(?:xlsx|pdf)$")
PM_RE = re.compile(r"^\d{3}\.\d{3}$")
PREFIX_SET = frozenset("CDGHLMNRST")
HG_SET = frozenset("DURLX")
FT_SET = frozenset("HIR")
ROUTE_CLAIM_RE = re.compile(r"\bRoute:\s*(\d{1,3}[A-Za-z]?)\b", re.IGNORECASE)
DIRECTION_RE = re.compile(r"\bDirection:\s*([NSEW])\s*[\u2013\u2014-]\s*([NSEW])\b")
DISTRICT_RE = re.compile(r"\bDistrict:\s*(\d{1,2})\b", re.IGNORECASE)
TRAILER_HEADING = "Unresolved Intersections"
LINE_TOLERANCE = 2.0
FRAGMENT_MAX_DISTANCE = 13.0

# Independent fixed-grid reconstruction.  Header anchors are validated on every
# data page, but they do not define these boundaries.  This is intentionally a
# different failure surface from production's per-page header-derived windows.
FIXED_WINDOWS = (
    ("County", -math.inf, 79.5),
    ("City", 79.5, 113.5),
    ("PM Prefix", 113.5, 136.5),
    ("PM", 136.5, 176.0),
    ("PM Suffix", 176.0, 197.0),
    ("HG", 197.0, 221.5),
    ("FT", 221.5, 244.5),
    ("Distance To Next Point", 244.5, 311.0),
    ("Description", 311.0, math.inf),
)
HEADER_TOKENS = ("COUNTY", "CITY", "PM", "HG", "FT", "NEXT", "POINT", "DESCRIPTION")
HEADER_ANCHOR_EXPECTED = {
    "COUNTY": 31.3, "CITY": 83.5, "PM": 149.1, "HG": 200.9,
    "FT": 225.3, "NEXT": 251.3, "POINT": 277.6, "DESCRIPTION": 317.2,
}
HEADER_ANCHOR_TOLERANCE = 12.0


class OracleError(RuntimeError):
    """The frozen source or the independent source contract was violated."""


@dataclass(frozen=True)
class FileEntry:
    name: str
    bytes: int
    sha256: str


@dataclass(frozen=True)
class SourceRow:
    source: str
    member: str
    route: str
    source_index: int
    source_ref: str
    page: int | None
    top: str | None
    values: tuple[str | None, ...]

    @property
    def identity_base(self) -> tuple[str, ...]:
        county = _county(self.values[0])
        prefix = _text(self.values[2])
        pm = _text(self.values[3])
        if pm:
            return self.route, county, prefix, pm
        return (
            self.route, county, "<PMLESS>", _text(self.values[5]),
            _text(self.values[6]), _pmless_kind(self.values[8]),
        )


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_bytes(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode("utf-8")


def _manifest(root: Path, suffix: str) -> tuple[dict[str, object], list[FileEntry]]:
    paths = sorted(root.glob(f"*{suffix}"), key=lambda path: path.name)
    entries = [FileEntry(path.name, path.stat().st_size, _sha_file(path)) for path in paths]
    wire = "".join(
        f"{entry.name}\t{entry.bytes}\t{entry.sha256}\n" for entry in entries
    ).encode("utf-8")
    return ({
        "files": len(entries),
        "bytes": sum(entry.bytes for entry in entries),
        "manifest_sha256": _sha_bytes(wire),
        "serialization": "name\\tbytes\\tsha256\\n sorted by name",
    }, entries)


def _bind_capture() -> dict[str, object]:
    observed_manifest = {
        "bytes": CAPTURE_MANIFEST.stat().st_size,
        "sha256": _sha_file(CAPTURE_MANIFEST),
    }
    if observed_manifest != {key: CAPTURE_MANIFEST_BINDING[key] for key in ("bytes", "sha256")}:
        raise OracleError(f"capture manifest identity drift: {observed_manifest}")
    capture = json.loads(CAPTURE_MANIFEST.read_text(encoding="utf-8"))
    if capture.get("stable_capture_sha256") != CAPTURE_MANIFEST_BINDING["stable_capture_sha256"]:
        raise OracleError("capture manifest stable identity drift")
    bound = {}
    for label, (files, size, digest, suffix) in TREE_BINDINGS.items():
        root = PRIVATE_ROOT / label
        observed, entries = _manifest(root, suffix)
        expected = {"files": files, "bytes": size, "manifest_sha256": digest}
        if any(observed[key] != value for key, value in expected.items()):
            raise OracleError(f"{label} private binding drift: {observed} != {expected}")
        bound[label] = {
            "root": str(root.resolve()), "observed": observed,
            "members": [asdict(entry) for entry in entries],
        }
    return {"capture_manifest": observed_manifest, "trees": bound}


def _route_from_member(path: Path) -> str:
    match = MEMBER_RE.fullmatch(path.name)
    if match is None:
        raise OracleError(f"unexpected Highway Sequence member: {path.name}")
    token = match.group(1).upper()
    return token[:3] + token[3:]


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _county(value: object) -> str:
    return _text(value).upper().rstrip(".")


def _display_text(value: object) -> str:
    text = re.sub(r"_x000d_", " ", _text(value), flags=re.IGNORECASE)
    text = text.replace("\r", " ").replace("\n", " ")
    return " ".join(text.split())


def _pmless_kind(value: object) -> str:
    text = _display_text(value).upper()
    for prefix in ("END OF ROUTE", "CITY END:", "COUNTY END:", "DISTRICT END:"):
        if text.startswith(prefix):
            return prefix
    return text


def _rows_digest(rows: Sequence[SourceRow]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        wire = _json_bytes({
            "source": row.source, "member": row.member, "route": row.route,
            "source_index": row.source_index, "source_ref": row.source_ref,
            "page": row.page, "top": row.top, "values": row.values,
        })
        digest.update(len(wire).to_bytes(8, "big"))
        digest.update(wire)
    return digest.hexdigest()


def _dataset_summary(label: str, rows: Sequence[SourceRow], extra: Mapping[str, object]) -> dict[str, object]:
    per_route = Counter(row.route for row in rows)
    nulls = {
        field: sum(row.values[index] in (None, "") for row in rows)
        for index, field in enumerate(FIELD_NAMES)
    }
    multiplicity = Counter(row.identity_base for row in rows)
    return {
        "label": label,
        "rows": len(rows),
        "columns": len(FIELD_NAMES),
        "ordered_raw_rows_sha256": _rows_digest(rows),
        "routes": len(per_route),
        "per_route_counts": dict(sorted(per_route.items())),
        "null_or_empty_by_field": nulls,
        "identity_base": {
            "unique": len(multiplicity),
            "duplicate_groups": sum(count > 1 for count in multiplicity.values()),
            "duplicate_occurrences": sum(count for count in multiplicity.values() if count > 1),
            "max_multiplicity": max(multiplicity.values(), default=0),
        },
        **dict(extra),
    }


def _excel_member(path: Path, source: str) -> tuple[list[SourceRow], dict[str, object]]:
    route = _route_from_member(path)
    workbook = load_workbook(
        path, read_only=True, data_only=False, keep_links=False,
    )
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            raise OracleError(f"{path.name}: sheet role universe {workbook.sheetnames}")
        sheet = workbook[SHEET_NAME]
        if sheet.sheet_state != "visible":
            raise OracleError(f"{path.name}: source sheet is not visible")
        if sheet.max_column != len(HEADERS):
            raise OracleError(f"{path.name}: max column {sheet.max_column}, expected 9")
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(HEADERS)))
        header = tuple(cell.value for cell in header_cells)
        if header != HEADERS:
            raise OracleError(f"{path.name}: positional header drift: {header!r}")
        rows: list[SourceRow] = []
        scalar_types: Counter[str] = Counter()
        for source_row, cells in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(HEADERS)), 2,
        ):
            values: list[str | None] = []
            for cell in cells:
                if cell.data_type in ("f", "e"):
                    raise OracleError(
                        f"{path.name}!{cell.coordinate}: formula/error is not source data"
                    )
                value = cell.value
                if value is None:
                    scalar_types["null"] += 1
                    values.append(None)
                elif isinstance(value, str):
                    scalar_types["str"] += 1
                    values.append(value)
                else:
                    raise OracleError(
                        f"{path.name}!{cell.coordinate}: unsupported source scalar "
                        f"{type(value).__name__}={value!r}"
                    )
            if all(value is None for value in values):
                raise OracleError(f"{path.name}: blank physical data row {source_row}")
            if not PM_RE.fullmatch(_text(values[3])) and _text(values[3]):
                raise OracleError(f"{path.name}: invalid PM at row {source_row}: {values[3]!r}")
            prefix = _text(values[2])
            suffix = _text(values[4])
            if prefix and prefix not in PREFIX_SET:
                raise OracleError(f"{path.name}: invalid prefix {prefix!r} at row {source_row}")
            if suffix not in ("", "E"):
                raise OracleError(f"{path.name}: invalid suffix {suffix!r} at row {source_row}")
            rows.append(SourceRow(
                source=source, member=path.name, route=route,
                source_index=source_row, source_ref=f"{path.name}:row:{source_row}",
                page=None, top=None, values=tuple(values),
            ))
        if sheet.max_row != len(rows) + 1:
            raise OracleError(
                f"{path.name}: physical row extent {sheet.max_row} != {len(rows) + 1}"
            )
        return rows, {
            "member": path.name, "route": route, "rows": len(rows),
            "max_row": sheet.max_row, "max_column": sheet.max_column,
            "scalar_types": dict(sorted(scalar_types.items())),
        }
    finally:
        workbook.close()


def _parse_excel_tree(root: Path, source: str) -> tuple[list[SourceRow], dict[str, object]]:
    rows: list[SourceRow] = []
    members = []
    for path in sorted(root.glob("*.xlsx"), key=lambda item: item.name):
        member_rows, diagnostic = _excel_member(path, source)
        rows.extend(member_rows)
        members.append(diagnostic)
    return rows, {
        "members": len(members),
        "member_diagnostics": members,
        "all_cells_string_or_null": all(
            set(item["scalar_types"]) <= {"str", "null"} for item in members
        ),
    }


def _cluster_lines(words: Sequence[Mapping[str, object]]) -> list[tuple[float, list[dict[str, object]]]]:
    groups: list[dict[str, object]] = []
    for source_word in sorted(words, key=lambda word: (float(word["top"]), float(word["x0"]))):
        word = dict(source_word)
        top = float(word["top"])
        candidates = []
        for index in range(max(0, len(groups) - 4), len(groups)):
            delta = abs(top - float(groups[index]["mean_top"]))
            if delta <= LINE_TOLERANCE:
                candidates.append((delta, index))
        if not candidates:
            groups.append({"mean_top": top, "tops": [top], "words": [word]})
        else:
            _, index = min(candidates)
            group = groups[index]
            group["tops"].append(top)
            group["mean_top"] = sum(group["tops"]) / len(group["tops"])
            group["words"].append(word)
    return [
        (float(group["mean_top"]), sorted(group["words"], key=lambda word: float(word["x0"])))
        for group in sorted(groups, key=lambda item: float(item["mean_top"]))
    ]


def _line_text(words: Sequence[Mapping[str, object]]) -> str:
    return " ".join(str(word["text"]) for word in words)


def _header_line(lines: Sequence[tuple[float, list[dict[str, object]]]]) -> tuple[float, dict[str, dict[str, object]]] | None:
    candidates = []
    for top, words in lines:
        texts = tuple(str(word["text"]) for word in words)
        positions = []
        cursor = 0
        for token in HEADER_TOKENS:
            try:
                index = texts.index(token, cursor)
            except ValueError:
                positions = []
                break
            positions.append(index)
            cursor = index + 1
        if positions:
            anchors = {token: words[index] for token, index in zip(HEADER_TOKENS, positions, strict=True)}
            candidates.append((top, anchors))
    if not candidates:
        return None
    if len(candidates) != 1:
        raise OracleError(f"data page has {len(candidates)} column-header candidates")
    top, anchors = candidates[0]
    for token, expected_x0 in HEADER_ANCHOR_EXPECTED.items():
        x0 = float(anchors[token]["x0"])
        if abs(x0 - expected_x0) > HEADER_ANCHOR_TOLERANCE:
            raise OracleError(f"header anchor {token} x0={x0:.3f}, expected {expected_x0:.3f}")
    return top, anchors


def _fixed_columns(words: Sequence[Mapping[str, object]]) -> dict[str, str]:
    tokens: dict[str, list[str]] = {field: [] for field in FIELD_NAMES}
    for word in words:
        center = (float(word["x0"]) + float(word["x1"])) / 2
        matches = [name for name, left, right in FIXED_WINDOWS if left <= center < right]
        if len(matches) != 1:
            raise OracleError(f"word has no unique fixed-grid owner: {word!r}")
        tokens[matches[0]].append(str(word["text"]))
    return {name: " ".join(tokens[name]) for name in FIELD_NAMES}


def _pmless(values: Mapping[str, str]) -> bool:
    return (
        not values["PM"] and not values["PM Prefix"] and not values["PM Suffix"]
        and bool(values["Description"])
        and values["HG"] in HG_SET and values["FT"] in FT_SET
    )


def _join_description(parts: Sequence[str]) -> str:
    result = ""
    for part in parts:
        if not result:
            result = part
        elif result.endswith("-"):
            result += part
        else:
            result += " " + part
    return result


def _pdf_member(path_string: str) -> tuple[list[dict[str, object]], dict[str, object]]:
    path = Path(path_string)
    route = _route_from_member(path)
    rows: list[SourceRow] = []
    page_diagnostics = []
    data_pages = 0
    trailer_seen = False
    fragment_count = 0
    with pdfplumber.open(path) as pdf:
        if len(pdf.pages) < 3:
            raise OracleError(f"{path.name}: expected cover, legend, and data pages")
        cover = " ".join((pdf.pages[0].extract_text() or "").split())
        legend = " ".join((pdf.pages[1].extract_text() or "").split())
        cover_claims = (
            "California Department of Transportation", "Highway Sequence Listing",
            "REPORT DATE : 7/9/2026", "REFERENCE DATE : 7/9/2026",
            f"ROUTE : {route}", "TASAS – TSMIS",
        )
        missing_cover = [claim for claim in cover_claims if claim not in cover]
        if missing_cover:
            raise OracleError(f"{path.name}: cover claims missing {missing_cover}")
        legend_claims = (
            "Legend", "Route Suffix Codes", "Post Mile Prefix Codes",
            "Post Mile Suffix Codes", "E - Equation", "Font Color",
        )
        missing_legend = [claim for claim in legend_claims if claim not in legend]
        if missing_legend:
            raise OracleError(f"{path.name}: legend claims missing {missing_legend}")

        for page_number, page in enumerate(pdf.pages[2:], 3):
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            lines = _cluster_lines(words)
            try:
                header = _header_line(lines)
            except OracleError as exc:
                raise OracleError(f"{path.name} p{page_number}: {exc}") from exc
            if header is None:
                raise OracleError(f"{path.name}: physical page {page_number} has no data header")
            data_pages += 1
            header_top, anchors = header
            header_bottom = max(float(word["bottom"]) for word in anchors.values())
            banner_text = " ".join(
                _line_text(line_words) for top, line_words in lines if top < header_top
            )
            route_claims = [_route_token(value) for value in ROUTE_CLAIM_RE.findall(banner_text)]
            if route_claims != [route]:
                raise OracleError(
                    f"{path.name} p{page_number}: route claims {route_claims}, expected {[route]}"
                )
            directions = [f"{left}-{right}" for left, right in DIRECTION_RE.findall(banner_text)]
            if len(directions) != 1:
                raise OracleError(f"{path.name} p{page_number}: direction claims {directions}")
            districts = [value.zfill(2) for value in DISTRICT_RE.findall(banner_text)]
            if len(districts) > 1:
                raise OracleError(f"{path.name} p{page_number}: district claims {districts}")

            page_rows: list[list[object]] = []
            fragments: list[tuple[float, str]] = []
            unclassified = []
            for top, line_words in lines:
                if top <= header_bottom + 2:
                    continue
                raw_text = _line_text(line_words)
                if raw_text.startswith(TRAILER_HEADING):
                    trailer_seen = True
                    break
                columns = _fixed_columns(line_words)
                pm = columns["PM"]
                if PM_RE.fullmatch(pm) or _pmless(columns):
                    prefix = columns["PM Prefix"]
                    suffix = columns["PM Suffix"]
                    if prefix and prefix not in PREFIX_SET:
                        raise OracleError(
                            f"{path.name} p{page_number}: unknown PM prefix {prefix!r}"
                        )
                    if suffix not in ("", "E"):
                        raise OracleError(
                            f"{path.name} p{page_number}: unknown PM suffix {suffix!r}"
                        )
                    if columns["HG"] and columns["HG"] not in HG_SET:
                        raise OracleError(
                            f"{path.name} p{page_number}: unknown HG {columns['HG']!r}"
                        )
                    if columns["FT"] and columns["FT"] not in FT_SET:
                        raise OracleError(
                            f"{path.name} p{page_number}: unknown FT {columns['FT']!r}"
                        )
                    values = [columns[field] or None for field in FIELD_NAMES]
                    page_rows.append([top, values, raw_text])
                elif columns["Description"] and not any(
                    columns[field] for field in FIELD_NAMES if field != "Description"
                ):
                    fragments.append((top, columns["Description"]))
                else:
                    unclassified.append({"top": f"{top:.3f}", "text": raw_text, "columns": columns})
            if unclassified:
                raise OracleError(
                    f"{path.name} p{page_number}: unclassified table lines: {unclassified[:3]}"
                )
            parts: dict[int, list[tuple[float, str]]] = {
                index: [(float(row[0]), _text(row[1][8]))]
                for index, row in enumerate(page_rows)
            }
            for fragment_top, fragment_text in fragments:
                distances = sorted(
                    (abs(float(row[0]) - fragment_top), index)
                    for index, row in enumerate(page_rows)
                )
                if not distances or distances[0][0] > FRAGMENT_MAX_DISTANCE:
                    raise OracleError(
                        f"{path.name} p{page_number}: orphan description fragment {fragment_text!r}"
                    )
                if len(distances) > 1 and abs(distances[0][0] - distances[1][0]) < 0.001:
                    raise OracleError(
                        f"{path.name} p{page_number}: ambiguous description fragment {fragment_text!r}"
                    )
                parts[distances[0][1]].append((fragment_top, fragment_text))
                fragment_count += 1
            for index, (top, values, _raw_text) in enumerate(page_rows):
                values[8] = _join_description(
                    [text for _part_top, text in sorted(parts[index]) if text]
                ) or None
                source_index = len(rows) + 1
                rows.append(SourceRow(
                    source="current_tsmis_pdf", member=path.name, route=route,
                    source_index=source_index,
                    source_ref=f"{path.name}:page:{page_number}:row:{index + 1}",
                    page=page_number, top=f"{float(top):.3f}", values=tuple(values),
                ))
            page_diagnostics.append({
                "page": page_number, "rows": len(page_rows), "fragments": len(fragments),
                "direction": directions[0], "district": districts[0] if districts else None,
                "header_x0": {
                    token: f"{float(word['x0']):.3f}" for token, word in anchors.items()
                },
            })
        metadata = {str(key): str(value) for key, value in sorted((pdf.metadata or {}).items())}
    return [asdict(row) for row in rows], {
        "member": path.name, "route": route, "pages": data_pages + 2,
        "data_pages": data_pages, "rows": len(rows), "fragments": fragment_count,
        "trailer_seen": trailer_seen, "metadata": metadata,
        "page_diagnostics": page_diagnostics,
    }


def _route_token(value: str) -> str:
    match = re.fullmatch(r"(\d{1,3})([A-Za-z]?)", value.strip())
    if match is None:
        raise OracleError(f"invalid route claim: {value!r}")
    return match.group(1).zfill(3) + match.group(2).upper()


def _parse_pdf_tree(root: Path, workers: int) -> tuple[list[SourceRow], dict[str, object]]:
    paths = sorted(root.glob("*.pdf"), key=lambda item: item.name)
    with ProcessPoolExecutor(max_workers=workers) as executor:
        results = list(executor.map(_pdf_member, (str(path) for path in paths), chunksize=1))
    rows: list[SourceRow] = []
    members = []
    for serialized_rows, diagnostic in results:
        rows.extend(SourceRow(**{**row, "values": tuple(row["values"])}) for row in serialized_rows)
        members.append(diagnostic)
    anchor_values: dict[str, list[float]] = defaultdict(list)
    for member in members:
        for page in member["page_diagnostics"]:
            for token, value in page["header_x0"].items():
                anchor_values[token].append(float(value))
    return rows, {
        "members": len(members),
        "pages": sum(item["pages"] for item in members),
        "data_pages": sum(item["data_pages"] for item in members),
        "description_fragments": sum(item["fragments"] for item in members),
        "members_with_trailer": sum(bool(item["trailer_seen"]) for item in members),
        "header_anchor_x0_ranges": {
            token: {"min": f"{min(values):.3f}", "max": f"{max(values):.3f}"}
            for token, values in sorted(anchor_values.items())
        },
        "member_diagnostics": members,
    }


def _indexed(rows: Sequence[SourceRow]) -> tuple[dict[tuple[str, ...], SourceRow], Counter[tuple[str, ...]]]:
    occurrences: Counter[tuple[str, ...]] = Counter()
    indexed = {}
    for row in rows:
        base = row.identity_base
        occurrences[base] += 1
        identity = (*base, f"occurrence:{occurrences[base]}")
        if identity in indexed:
            raise OracleError(f"duplicate occurrence-qualified identity: {identity}")
        indexed[identity] = row
    return indexed, occurrences


def _canonical_values(row: SourceRow, *, display: bool) -> tuple[str, ...]:
    values = []
    for index, value in enumerate(row.values):
        if index == 0:
            values.append(_county(value))
        elif display:
            values.append(_display_text(value))
        else:
            values.append(_text(value))
    return tuple(values)


def _parity(left_rows: Sequence[SourceRow], right_rows: Sequence[SourceRow], label: str) -> dict[str, object]:
    left, left_multiplicity = _indexed(left_rows)
    right, right_multiplicity = _indexed(right_rows)
    shared = sorted(set(left) & set(right))
    left_only = sorted(set(left) - set(right))
    right_only = sorted(set(right) - set(left))
    differing = []
    raw_exact = display_exact = 0
    field_differences: Counter[str] = Counter()
    display_field_differences: Counter[str] = Counter()
    for identity in shared:
        left_row, right_row = left[identity], right[identity]
        left_raw = _canonical_values(left_row, display=False)
        right_raw = _canonical_values(right_row, display=False)
        left_display = _canonical_values(left_row, display=True)
        right_display = _canonical_values(right_row, display=True)
        raw_fields = [
            field for index, field in enumerate(FIELD_NAMES)
            if left_raw[index] != right_raw[index]
        ]
        display_fields = [
            field for index, field in enumerate(FIELD_NAMES)
            if left_display[index] != right_display[index]
        ]
        if not raw_fields:
            raw_exact += 1
        if not display_fields:
            display_exact += 1
        if raw_fields:
            field_differences.update(raw_fields)
            display_field_differences.update(display_fields)
            differing.append({
                "identity": identity,
                "left_ref": left_row.source_ref,
                "right_ref": right_row.source_ref,
                "left_values": left_row.values,
                "right_values": right_row.values,
                "raw_differing_fields": raw_fields,
                "display_differing_fields": display_fields,
            })
    return {
        "label": label,
        "identity_policy": (
            "route + normalized county + prefix + PM + occurrence; PM-less rows use "
            "route/county/HG/FT/description-kind + occurrence"
        ),
        "left_rows": len(left_rows), "right_rows": len(right_rows),
        "paired_rows": len(shared), "left_only_rows": len(left_only),
        "right_only_rows": len(right_only), "raw_exact_rows": raw_exact,
        "raw_differing_rows": len(shared) - raw_exact,
        "display_exact_rows": display_exact,
        "display_differing_rows": len(shared) - display_exact,
        "raw_field_difference_counts": dict(sorted(field_differences.items())),
        "display_field_difference_counts": dict(sorted(display_field_differences.items())),
        "left_only": [
            {"identity": identity, "ref": left[identity].source_ref, "values": left[identity].values}
            for identity in left_only
        ],
        "right_only": [
            {"identity": identity, "ref": right[identity].source_ref, "values": right[identity].values}
            for identity in right_only
        ],
        "differing_pairs": differing,
        "multiplicity_base_differences": [
            {"identity_base": identity, "left": left_multiplicity[identity], "right": right_multiplicity[identity]}
            for identity in sorted(set(left_multiplicity) | set(right_multiplicity))
            if left_multiplicity[identity] != right_multiplicity[identity]
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--row-cache", type=Path, default=DEFAULT_ROW_CACHE)
    parser.add_argument("--workers", type=int, default=min(8, os.cpu_count() or 1))
    args = parser.parse_args()
    if args.workers < 1:
        raise OracleError("workers must be positive")

    binding = _bind_capture()
    current_excel_rows, current_excel_diag = _parse_excel_tree(
        PRIVATE_ROOT / "current_tsmis_excel", "current_tsmis_excel",
    )
    historical_excel_rows, historical_excel_diag = _parse_excel_tree(
        PRIVATE_ROOT / "historical_tsmis_excel_7_8", "historical_tsmis_excel_7_8",
    )
    pdf_rows, pdf_diag = _parse_pdf_tree(PRIVATE_ROOT / "current_tsmis_pdf", args.workers)

    current_parity = _parity(
        current_excel_rows, pdf_rows, "current July-9 TSMIS Excel vs current July-9 TSMIS PDF",
    )
    historical_parity = _parity(
        historical_excel_rows, pdf_rows, "historical July-8 TSMIS Excel vs July-9 TSMIS PDF",
    )
    edition_delta = _parity(
        historical_excel_rows, current_excel_rows, "historical July-8 Excel vs current July-9 Excel",
    )
    extraction_invariants = {
        "current_excel_members_252": current_excel_diag["members"] == 252,
        "historical_excel_members_252": historical_excel_diag["members"] == 252,
        "current_pdf_members_252": pdf_diag["members"] == 252,
        "current_excel_rows_60494": len(current_excel_rows) == 60_494,
        "historical_excel_rows_60493": len(historical_excel_rows) == 60_493,
        "current_pdf_rows_60493": len(pdf_rows) == 60_493,
        "excel_cells_string_or_null": (
            current_excel_diag["all_cells_string_or_null"]
            and historical_excel_diag["all_cells_string_or_null"]
        ),
        "all_pdf_pages_classified": pdf_diag["data_pages"] == pdf_diag["pages"] - 504,
    }
    result = {
        "audit": "Stage 8 Highway Sequence independent TSMIS source oracle draft",
        "status": "source-parsed" if all(extraction_invariants.values()) else "failed",
        "audit_complete": False,
        "reason_not_complete": (
            "Every residual below must be semantically classified and then projected "
            "independently against authoritative raw/normalized TSN before acceptance."
        ),
        "binding": binding,
        "datasets": {
            "current_tsmis_excel": _dataset_summary(
                "current_tsmis_excel", current_excel_rows, current_excel_diag,
            ),
            "historical_tsmis_excel_7_8": _dataset_summary(
                "historical_tsmis_excel_7_8", historical_excel_rows, historical_excel_diag,
            ),
            "current_tsmis_pdf": _dataset_summary(
                "current_tsmis_pdf", pdf_rows, pdf_diag,
            ),
        },
        "extraction_invariants": extraction_invariants,
        "source_parity": {
            "current_excel_vs_pdf": current_parity,
            "historical_excel_vs_pdf": historical_parity,
            "historical_vs_current_excel": edition_delta,
        },
    }
    if result["status"] != "source-parsed":
        raise OracleError(f"source extraction invariants failed: {extraction_invariants}")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_bytes(_json_bytes(result))
    row_cache = {
        "audit": "Highway Sequence independent source-row development cache",
        "not_an_acceptance_artifact": True,
        "capture_manifest": binding["capture_manifest"],
        "dataset_digests": {
            "current_tsmis_excel": result["datasets"]["current_tsmis_excel"]["ordered_raw_rows_sha256"],
            "historical_tsmis_excel_7_8": result["datasets"]["historical_tsmis_excel_7_8"]["ordered_raw_rows_sha256"],
            "current_tsmis_pdf": result["datasets"]["current_tsmis_pdf"]["ordered_raw_rows_sha256"],
        },
        "rows": {
            "current_tsmis_excel": [asdict(row) for row in current_excel_rows],
            "historical_tsmis_excel_7_8": [asdict(row) for row in historical_excel_rows],
            "current_tsmis_pdf": [asdict(row) for row in pdf_rows],
        },
    }
    args.row_cache.parent.mkdir(parents=True, exist_ok=True)
    args.row_cache.write_bytes(_json_bytes(row_cache))
    print(
        "PASS Highway Sequence independent source parse: "
        f"current Excel {len(current_excel_rows):,}; PDF {len(pdf_rows):,}; "
        f"historical Excel {len(historical_excel_rows):,}; "
        f"current paired {current_parity['paired_rows']:,}, "
        f"one-sided {current_parity['left_only_rows']}/{current_parity['right_only_rows']}, "
        f"different {current_parity['raw_differing_rows']:,}; output {args.output}; "
        f"development rows {args.row_cache}"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except OracleError as exc:
        print(f"FAIL Highway Sequence independent source oracle: {exc}")
        raise SystemExit(1)
