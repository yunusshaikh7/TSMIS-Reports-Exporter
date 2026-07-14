"""E2 build-time identity/pairing freshness gate.

Default mode proves the snapshot/check formula structure without Excel.
``--excel`` performs adversarial edits and requires every static-identity change
to replace a formerly green banner with ``REGENERATE REQUIRED``.
"""
from __future__ import annotations

import argparse
from pathlib import Path
import shutil
import sys

from _checklib import Checker, scripts_path, temp_dir

scripts_path()

from compare_core import CompareSchema, run_compare  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


ROWS = [
    ["UNIQUE", "same"],
    ["DUP", "left"],
    ["DUP", "right"],
]
SCHEMA = CompareSchema(
    report_name="E2 Build Freshness", header=["Key", "Value"],
    side_a="A", side_b="B", id_noun="row", id_noun_plural="rows")


def _header_col(ws, label):
    positions = [cell.column for cell in ws[1] if cell.value == label]
    if len(positions) != 1:
        raise AssertionError((label, positions))
    return positions[0]


def _excel_edit_and_rebuild(excel, source, target, edit):
    shutil.copy2(source, target)
    book = None
    try:
        book = excel.Workbooks.Open(
            str(Path(target).resolve()), UpdateLinks=0, ReadOnly=False)
        edit(book)
        excel.CalculateFullRebuild()
        book.Save()
        book.Close(SaveChanges=False)
        book = None
    finally:
        if book is not None:
            book.Close(SaveChanges=False)


def _calculated_state(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        summary = wb["Summary"]
        banner = summary["B3"].value
        freshness = None
        for row in summary.iter_rows(values_only=True):
            if (len(row) > 2 and row[1]
                    == "Build-time source identity and duplicate pairing snapshot is current"):
                freshness = row[2]
        comparison = wb["Comparison"]
        header = list(next(comparison.iter_rows(values_only=True)))
        diffs_col = header.index("Diffs")
        diffs = [row[diffs_col] for row in comparison.iter_rows(
            min_row=2, values_only=True)]
        return banner, freshness, diffs
    finally:
        wb.close()


def test_structure(c, path):
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        snapshots = ["__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B"]
        c.check("both immutable source snapshots exist and are very hidden",
                all(name in wb.sheetnames
                    and wb[name].sheet_state == "veryHidden"
                    for name in snapshots))
        failures = []
        for side, snapshot in (("A", snapshots[0]), ("B", snapshots[1])):
            ws = wb[side]
            headers = [cell.value for cell in ws[1]]
            fresh_cols = [index + 1 for index, value in enumerate(headers)
                          if isinstance(value, str)
                          and value.startswith("__CMP_E2_BUILD_FRESH_V1_")]
            if not fresh_cols:
                failures.append((side, "no freshness columns"))
                continue
            formulas = [ws.cell(row, fresh_cols[0]).value
                        for row in range(2, len(ROWS) + 2)]
            if not all(isinstance(value, str) and value.startswith("=IF(AND(")
                       and snapshot in value and '"OK","STALE"' in value
                       for value in formulas):
                failures.append((side, formulas))
            if not all(ws.column_dimensions[ws.cell(1, col).column_letter].hidden
                       for col in fresh_cols):
                failures.append((side, "visible freshness column"))
            key_col = _header_col(ws, "Key (helper)")
            snap = wb[snapshot]
            for row in range(2, len(ROWS) + 2):
                if ws.cell(row, key_col).value != snap.cell(row, key_col).value:
                    failures.append((side, row, "helper snapshot mismatch"))
                if tuple(ws.cell(row, col).value for col in (2, 3)) \
                        != tuple(snap.cell(row, col).value for col in (2, 3)):
                    failures.append((side, row, "source snapshot mismatch"))
        c.check("data-sheet chunks compare exact source/helper cells to snapshots",
                not failures, repr(failures[:5]))

        summary = wb["Summary"]
        summary_formulas = [cell.value for row in summary.iter_rows()
                            for cell in row if cell.data_type == "f"]
        banner_formula = summary["B3"].value
        c.check("Summary verdict fails closed on any stale snapshot chunk",
                isinstance(banner_formula, str)
                and "REGENERATE REQUIRED" in banner_formula
                and all(name in banner_formula for name in snapshots),
                repr(banner_formula))
        c.check("Summary has an explicit freshness self-check",
                any("Build-time source identity and duplicate pairing snapshot is current"
                    == row[1].value for row in summary.iter_rows()
                    if len(row) > 1)
                and any('"STALE"' in str(formula) for formula in summary_formulas))
    finally:
        wb.close()


def test_excel(c, source, root):
    print("\nInstalled Excel build-freshness edits:")
    scenarios = (
        ("baseline", lambda _book: None, False, None),
        ("ordinary value edit",
         lambda book: setattr(book.Worksheets("A").Range("C2"), "Value", "changed"),
         True, None),
        ("identity-key edit",
         lambda book: setattr(book.Worksheets("A").Range("B2"), "Value", "NEWKEY"),
         True, None),
        ("opaque helper edit",
         lambda book: setattr(book.Worksheets("A").Range("D2"), "Value", "tampered"),
         True, None),
        ("duplicate assignment made stale",
         lambda book: (
             setattr(book.Worksheets("B").Range("C3"), "Value", "right"),
             setattr(book.Worksheets("B").Range("C4"), "Value", "left")),
         True, [0, 1, 1]),
        ("source row deleted",
         lambda book: book.Worksheets("A").Rows(2).Delete(),
         True, None),
    )
    import win32com.client as win32
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        for index, (name, edit, stale, expected_diffs) in enumerate(scenarios):
            target = root / f"excel-{index}.xlsx"
            try:
                _excel_edit_and_rebuild(excel, source, target, edit)
                banner, freshness, diffs = _calculated_state(target)
                error = None
            except Exception as exc:
                banner = freshness = diffs = None
                error = f"{type(exc).__name__}: {exc}"
            expected_banner = (isinstance(banner, str)
                               and ("REGENERATE REQUIRED" in banner) == stale
                               and (not stale or banner.startswith("✗")))
            c.check(f"{name}: certification state is correct",
                    error is None and expected_banner
                    and freshness == ("CHECK" if stale else "OK"),
                    f"error={error!r}; banner={banner!r}; fresh={freshness!r}")
            if expected_diffs is not None:
                c.check(
                    f"{name}: fixed pairing can show phantom observations but not certify",
                    diffs == expected_diffs,
                    f"expected={expected_diffs!r}; actual={diffs!r}")
    finally:
        if excel is not None:
            excel.Quit()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", action="store_true")
    args = parser.parse_args()
    c = Checker()
    with temp_dir("tsmis_e2_build_freshness_") as root:
        source = root / "freshness.xlsx"
        result = run_compare(
            SCHEMA, ROWS, ROWS, False, source, mode="formulas")
        c.check("freshness fixture builds", result.status == "ok", repr(result))
        if result.status == "ok":
            test_structure(c, source)
            if args.excel:
                test_excel(c, source, root)
    return c.summary()


if __name__ == "__main__":
    raise SystemExit(main())
