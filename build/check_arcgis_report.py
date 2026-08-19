"""Golden check for the ArcGIS "Reports vs layers" lane (v0.39.0): the CA
HIGHWAYS build PROJECTED onto a report's own shape, and the TSMIS-vs-TSMIS
comparison over it.

The two rules that make the projection a BUILD rather than a column rename are
the ones with teeth, so they are driven from a synthetic CA HIGHWAYS workbook
end to end through the shipped `consolidate()`:

  * adjacent spans that agree across every PRINTED column merge into one record,
    and the printed Length is the merged span's own extent (a THY split on a
    column Highway Detail never shows must not become a record boundary);
  * Description is START-ANCHORED — a following blank continues the record, a
    following DIFFERENT description starts a new one.

Plus the contracts the comparison depends on: the emitted header is exactly the
consolidated export's (so the shared loader reads it), `RU Eff` is declared a
context column on both sides, and the role gates refuse a swapped pair.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_arcgis_report.py
"""
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import Workbook, load_workbook

import arcgis_report_highway_detail as ah
import clean_highway_columns as chc
import compare_highway_detail_arcgis as cmp_arc
import consolidate_clean_highway as cch
import highway_detail_columns as hdc
import outcome

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _raises(fn, needle=None):
    try:
        fn()
        return False
    except ValueError as e:
        return needle is None or needle.lower() in str(e).lower()


_T = {n: i for i, n in enumerate(chc.ARC_HEADER)}


def thy(begin, end, **over):
    """One CA HIGHWAYS row: a plausible base record with `over` applied."""
    r = [None] * len(chc.ARC_HEADER)
    base = {
        "THY_COUNTY_CODE": "ORA", "THY_ROUTE_NAME": "001",
        "THY_PM_PREFIX_CODE": "R", "THY_BEGIN_PM_AMT": begin,
        "THY_END_PM_AMT": end, "THY_LENGTH_MILES_AMT": end - begin,
        "THY_RECORD_DATE": "73-10-19", "THY_HIGHWAY_GROUP_CODE": "D",
        "THY_HIGHWAY_ACCESS_CODE": "F", "THY_ACCESS_EFF_DATE": "73-10-19",
        "THY_CITY_CODE": "DAPT", "THY_POPULATION_CODE": "U",
        "THY_NON_ADD_CODE": "A", "THY_LT_LANES_AMT": 3,
        "THY_MEDIAN_WIDTH_AMT": 14, "THY_MEDIAN_WIDTH_VAR_CODE": "Z",
    }
    base.update(over)
    for k, v in base.items():
        r[_T[k]] = v
    return r


def build_source(path, rows, skips=None):
    """A CA HIGHWAYS build. `skips` adds the HF-01 unassertable-span record the
    real build writes when it could not place a source span."""
    wb = Workbook()
    ws = wb.active
    ws.title = chc.ARC_SHEET
    ws.append(list(chc.ARC_HEADER))
    for r in rows:
        ws.append(r)
    mk = wb.create_sheet(chc.ARC_MARKER_SHEET)
    mk.append(["Build version", 1])
    mk.append(["As-of date", "2026-08-17"])
    mk.append(["Layer library", r"C:\demo\arcgis_layers"])
    if skips:
        spans, marked = skips
        mk.append(["Skipped source spans", spans])
        mk.append(["Marked anchor cells", marked])
        mk.append(["Unavailable marker", cch.UNAVAILABLE_TOKEN])
        mk.append(["Skipped source reason", cch.SKIP_REASON])
    wb.save(path)


def read_out(path):
    wb = load_workbook(path, read_only=True)
    ws = wb[ah.SHEET_NAME]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) + [None] * (len(header) - len(r))
            for r in it if any(c is not None for c in r)]
    wb.close()
    return header, rows


def main():
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_agrep_"))
    src, out = tmp / "clean_highway_built.xlsx", tmp / "hd_from_layers.xlsx"

    print("the projection contract:")
    check("every printed column is mapped (import-time assert holds)",
          set(ah.PROJECTION) == set(hdc.HEADER))
    check("every printed column now has a source — no context columns left "
          "(DA2 closed: RU Eff prints THY_POPULATION_EFF_DATE)",
          ah.CONTEXT_COLUMNS == () and cmp_arc.CONTEXT_FIELDS == ())
    check("RU Eff maps to the build-only population effective date",
          ah.PROJECTION["RU Eff"][0] == ("THY_POPULATION_EFF_DATE",))
    check("...which the TSN 74-column schema does NOT carry, so the TSN raw "
          "gate is untouched and our build is a strict superset",
          "THY_POPULATION_EFF_DATE" not in chc.HEADER
          and list(chc.ARC_HEADER[:len(chc.HEADER)]) == list(chc.HEADER)
          and chc.ARC_HEADER[-1] == "THY_POPULATION_EFF_DATE")
    check("...and it is CONTEXT on the vs-TSN side, where TSN has no column "
          "to compare it against",
          "THY_POPULATION_EFF_DATE" in chc.CONTEXT_COLUMNS)
    check("the merge rule ignores position AND description, nothing else",
          set(hdc.HEADER) - set(ah._MERGE_FIELDS)
          == {"Post Mile", "Length", "Description"})

    print("the merge — a THY split on an unprinted column is not a record:")
    # Two spans identical on every PRINTED column (they differ only in ADT, which
    # Highway Detail never shows) must become ONE record of the merged extent.
    build_source(src, [thy(0.129, 0.170, THY_ADT_AMT=1000,
                           THY_LANDMARK_SHORT_DESC="JCT 5 CAMINO"),
                       thy(0.170, 0.204, THY_ADT_AMT=2000)])
    res = ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    header, rows = read_out(out)
    check("build ok", res.status == "ok")
    check("the emitted header IS the consolidated export's (Route + the 34 "
          "labels) — the shared loader reads it unchanged",
          header == [hdc.ROUTE_COL] + list(hdc.HEADER))
    check("the two spans became ONE record", len(rows) == 1)
    pm_i = header.index("Post Mile")
    len_i = header.index("Length")
    desc_i = header.index("Description")
    check("...starting at the FIRST span's postmile", rows[0][pm_i] == "R000.129")
    check("...with the MERGED extent as the printed Length (0.204-0.129), not "
          "the first span's own 0.041", rows[0][len_i] == "000.075")
    check("...keeping the start-anchored description",
          rows[0][desc_i] == "JCT 5 CAMINO")
    check("the result says how many rows it merged away",
          "merged" in (res.message or "").lower())

    print("the merge STOPS where the report would print a new record:")
    # A printed column differs -> two records.
    build_source(src, [thy(0.129, 0.170, THY_LT_LANES_AMT=3),
                       thy(0.170, 0.204, THY_LT_LANES_AMT=2)])
    ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    _h, rows = read_out(out)
    check("a differing PRINTED column blocks the merge", len(rows) == 2)

    # A DIFFERENT non-blank description starts a record (a landmark is exactly
    # what does), while a blank one continues it.
    build_source(src, [thy(0.129, 0.170, THY_LANDMARK_SHORT_DESC="JCT 5"),
                       thy(0.170, 0.204, THY_LANDMARK_SHORT_DESC="RAMP NOSE")])
    ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    _h, rows = read_out(out)
    check("a DIFFERENT description starts a new record", len(rows) == 2)

    # Non-touching spans never merge, however equal.
    build_source(src, [thy(0.129, 0.170), thy(0.200, 0.240)])
    ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    _h, rows = read_out(out)
    check("a GAP between two identical spans keeps them separate", len(rows) == 2)

    print("cell rendering matches the report's print forms:")
    build_source(src, [thy(0.129, 0.204, THY_LT_LANES_AMT=3,
                           THY_MEDIAN_WIDTH_AMT=14,
                           THY_MEDIAN_WIDTH_VAR_CODE="Z",
                           THY_NON_ADD_CODE="A",
                           THY_LANDMARK_SHORT_DESC="jct 5 camino")])
    ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    header, rows = read_out(out)
    at = lambda name: rows[0][header.index(name)]
    check("lane counts zero-pad to 2 digits", at("LB #Ln") == "03")
    check("median width + variance glue ('14Z')", at("Med V/WDA") == "14Z")
    check("'A' (add mileage) prints BLANK, as the report does",
          at("NA") in (None, ""))
    check("the description prints upper-case", at("Description") == "JCT 5 CAMINO")

    print("DA2 — RU Eff prints the population block's own effective date:")
    import datetime
    build_source(src, [thy(0.129, 0.204, THY_POPULATION_CODE="U",
                           THY_POPULATION_EFF_DATE=datetime.date(1964, 1, 1))])
    ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    header, rows = read_out(out)
    at2 = lambda name: rows[0][header.index(name)]
    check("a real date renders in the report's YY-MM-DD form, not blank",
          at2("RU Eff") == "64-01-01")
    check("...beside the code it belongs to", at2("RU") == "U")
    # It is a printed column like any other, so a change in it is a record
    # boundary — the same rule every other printed column obeys.
    build_source(src, [thy(0.129, 0.170,
                           THY_POPULATION_EFF_DATE=datetime.date(1964, 1, 1)),
                       thy(0.170, 0.204,
                           THY_POPULATION_EFF_DATE=datetime.date(2010, 12, 31))])
    ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    _h, rows = read_out(out)
    check("a differing RU Eff now BLOCKS the merge, as a printed column must",
          len(rows) == 2)
    # A span the population layer never covered has no date, and that must read
    # as blank rather than as some neighbour's date.
    build_source(src, [thy(0.129, 0.204)])
    ah.consolidate(built_path=src, out_path=out, confirm_overwrite=lambda p: True)
    header, rows = read_out(out)
    check("a span with no population date renders blank, never invented",
          rows[0][header.index("RU Eff")] in (None, ""))

    print("a PARTIAL source build stays partial through the projection:")
    # HF-01/RB-1: where the CA HIGHWAYS build could not place a source span it
    # writes a reserved token into the anchor cell. The token travels into this
    # report, so the projection must not report COMPLETE over it and the
    # comparison must not COUNT it as a difference — it says "unknowable", not
    # "different".
    build_source(src, [thy(0.129, 0.204,
                           THY_LT_LANES_AMT=cch.UNAVAILABLE_TOKEN)],
                 skips=(102, 174))
    res = ah.consolidate(built_path=src, out_path=out,
                         confirm_overwrite=lambda p: True)
    check("the projection still builds", res.status == "ok")
    check("...but reports PARTIAL, not COMPLETE, over an unassertable source",
          res.completion == outcome.PARTIAL)
    check("...and carries the source's skipped-span count",
          res.skipped_inputs == 102)
    check("...and says so in the message",
          "102" in (res.message or "") and "assert" in (res.message or ""))
    facts = ah.report_facts(out)
    check("the report's own marker sheet carries the counts forward",
          facts.get("skipped_source_spans") == "102"
          and facts.get("marked_anchor_cells") == "174"
          and facts.get("unavailable_marker") == cch.UNAVAILABLE_TOKEN)
    rule = cmp_arc._unavailable_rule(out)
    check("the comparison arms the non-asserting rule off those facts",
          bool(rule) and rule[0] == cch.UNAVAILABLE_TOKEN)
    check("...and its Summary note states both counts",
          bool(rule) and "102" in rule[1] and "174" in rule[1])
    check("the marker cell reached the report (the fixture has teeth)",
          any(cch.UNAVAILABLE_TOKEN in str(c)
              for c in read_out(out)[1][0]))

    print("a skip-free build is untouched:")
    build_source(src, [thy(0.129, 0.204)])
    res = ah.consolidate(built_path=src, out_path=out,
                         confirm_overwrite=lambda p: True)
    check("a build with no skips is COMPLETE", res.completion == outcome.COMPLETE)
    check("...with no skipped inputs", not res.skipped_inputs)
    check("...and the comparison arms no rule at all",
          cmp_arc._unavailable_rule(out) == ())

    print("identity + role gates:")
    check("our build identifies itself", ah.is_arcgis_report(out))
    facts = ah.report_facts(out)
    check("...and carries the SOURCE build's as-of date, not today's",
          facts.get("asof") == "2026-08-17")
    # A TSMIS-shaped workbook without our marker sheet is not ours.
    plain = tmp / "tsmis_like.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = ah.SHEET_NAME
    ws.append([hdc.ROUTE_COL] + list(hdc.HEADER))
    ws.append(["001", "R000.129", "000.075"] + [""] * (len(hdc.HEADER) - 3))
    wb.save(plain)
    check("a TSMIS-shaped workbook without the marker is NOT ours",
          not ah.is_arcgis_report(plain))
    check("an unreadable/missing path answers False rather than raising",
          not ah.is_arcgis_report(tmp / "nope.xlsx"))
    check("the comparison refuses a NON-ArcGIS side A",
          _raises(lambda: cmp_arc._load_pair(plain, plain), "not an ArcGIS-built"))
    check("...and refuses an ArcGIS side B (a swapped pair)",
          _raises(lambda: cmp_arc._load_pair(out, out), "not a TSMIS export"))

    print("the projection refuses a source that is not the CA HIGHWAYS build:")
    check("a missing source names the Clean Road build",
          ah.consolidate(built_path=tmp / "absent.xlsx", out_path=out,
                         confirm_overwrite=lambda p: True).status == "error")
    wrong = tmp / "wrong.xlsx"
    wb = Workbook()
    wb.active.title = "Not It"
    wb.active.append(["a", "b"])
    wb.save(wrong)
    check("a workbook with no CA HIGHWAYS sheet is refused",
          ah.consolidate(built_path=wrong, out_path=out,
                         confirm_overwrite=lambda p: True).status == "error")

    print()
    if _fail:
        print(f"FAILED {len(_fail)} check(s):")
        for f in _fail:
            print(f"  - {f}")
        return 1
    print("ALL ARCGIS-REPORT CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
