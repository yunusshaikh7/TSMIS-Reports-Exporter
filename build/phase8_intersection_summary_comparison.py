#!/usr/bin/env python3
"""Independent Stage-8 Intersection Summary raw-to-comparison oracle.

Authority flows from the exact 217-route All Reports 7.9 TSMIS Excel/PDF pair,
the exact accepted TSN raw-PDF -> r7-normalized chain, and the authoritative
TSNR control/geometry crosswalk.  Truth-side code imports no application parser,
normalizer, schema, comparator, or writer.  Production executes only in a child
process after source truth is complete, and its emitted workbooks are parsed back
against that independently derived truth.

The outcome flags deliberately answer different questions:

* ``source_truth_exact``: every bound source, route, category, value, provenance
  field, cross-format relation, and TSN decision source is exact.
* ``production_value_projection_exact``: the current product emits the exact
  accepted current-source values.
* ``production_comparison_semantics_exact``: the product's current-source union,
  one-sided taxonomy, statuses, and verdict inputs match the approved base truth.
* ``normalized_source_full_conservation``: raw TSN source claims can be rebuilt
  from normalized bytes.  This remains false while J/K/L/M/N/P, the erroneous
  raw F label, and printed provenance are not retained.
* ``comparison_end_to_end_perfect``: raw-source conservation plus fail-closed
  production behavior and durable provenance are all complete.  Known product
  findings keep this false until remediation.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
import hashlib
import json
import logging
import os
from pathlib import Path
import re
import subprocess
import sys
import tempfile
from typing import Iterable, Sequence
import zipfile

import openpyxl
from openpyxl import load_workbook
import pdfplumber


# The authoritative PDFs contain a recoverable missing FontBBox entry. pdfminer logs
# the same message repeatedly while still extracting every bound word correctly. The
# all-member pixel census and cross-render review independently cover visual integrity;
# suppress only this parser-library diagnostic noise so structured audit output remains
# usable.
logging.getLogger("pdfminer").setLevel(logging.ERROR)


REPO_ROOT = Path(__file__).resolve().parent.parent
GENERATOR_PATH = Path(__file__).resolve()
SELF_GATE_PATH = GENERATOR_PATH.with_name(
    "check_phase8_intersection_summary_comparison.py")
PRODUCT_HELPER_PATH = GENERATOR_PATH.with_name(
    "phase8_intersection_summary_product_witness.py")

SOURCE_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9"
    r"\2026-07-09 ars-prod")
DEFAULT_SUMMARY_XLSX_ROOT = SOURCE_ROOT / "intersection_summary"
DEFAULT_SUMMARY_PDF_ROOT = SOURCE_ROOT / "intersection_summary_pdf"
DEFAULT_TSN_RAW = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\intersection_summary\raw"
    r"\Intersection Summary Statewide_TSN.pdf")
REFERENCE_ROOT = Path(r"C:\Users\Yunus\Downloads\TSMIS\reference")
DEFAULT_TSNR_REFERENCE = (
    REFERENCE_ROOT / "TSNR - Intersection Control and Geometry Type_4.25.24_AT 1.xlsx")
VISUAL_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd")
R7_ROOT = (
    VISUAL_ROOT / "phase4_tsn_rebaseline" / "raw-2026-07-12-r7" /
    "intersection_summary" / "consolidated")
DEFAULT_TSN_XLSX = R7_ROOT / "tsn_intersection_summary_normalized.xlsx"
DEFAULT_STAGE6_RESULT = (
    VISUAL_ROOT / "phase6_tsn_conservation" /
    "intersection_summary_conservation_r7.json")
DEFAULT_STAGE6_ACCEPTANCE = Path(str(DEFAULT_STAGE6_RESULT) + ".acceptance.json")
DEFAULT_CROSS_FORMAT_RESULT = (
    VISUAL_ROOT / "phase4_tsn_rebaseline" /
    "intersection-tsn-cross-format-oracle-v2.json")
DEFAULT_WORK_ROOT = REPO_ROOT / "tmp" / "phase8-intersection-summary-oracle"


TREE_BINDINGS = {
    "tsmis_xlsx": {
        "files": 217,
        "bytes": 5_953_364,
        "manifest_sha256": "e3e235e0f48645750b65b9df966a963c5a9bb856798d23661c95ab44056956e5",
        "suffix": ".xlsx",
    },
    "tsmis_pdf": {
        "files": 217,
        "bytes": 21_518_480,
        "manifest_sha256": "63f06f7b7f483a1fcd85be60278e7eebfbab51a79a1de955e9d3eac5bb8c8c2a",
        "suffix": ".pdf",
    },
}

FILE_BINDINGS = {
    "tsn_raw_pdf": {
        "bytes": 12_326,
        "sha256": "c3ad85848764df1b6da53c0bba0f785b3c045e83675f5983555ef514688a7d46",
    },
    "tsn_normalized_xlsx": {
        "bytes": 6_323,
        "sha256": "94befb313416a356a6e9f0363ffae0d065bd03c15ea1fce5bd8e93e0bf59a210",
    },
    "stage6_result": {
        "bytes": 245_040,
        "sha256": "f3a0aa0dfb15cf2ca911ec98721c8dcc0d5d9b25c0ce3cc89184d2959aaf64de",
    },
    "stage6_acceptance": {
        "bytes": 44_337,
        "sha256": "cdf63defdb62d2066a2cafb7229d0c1539a0c6d90f80ea1b96c07c77f609b703",
    },
    "tsn_cross_format_result": {
        "bytes": 91_032,
        "sha256": "63f5741203b06ef37245f195953058cf45ec921c04aaa00ccf676e44baba2c2e",
    },
    "tsnr_reference": {
        "bytes": 15_419,
        "sha256": "64140ca7ef38b1d06c2a8112b99d9f327b3812d6c399c1eb417b338dc59db23e",
    },
}

EXPECTED_ROUTE_COUNT = 217
EXPECTED_ROUTE_LF_SHA256 = (
    "0dcd88a8b8f8156a87c7cc7834972aa08b018f5c36f03fa469b5750236b01a8d")
EXPECTED_SUFFIX_ROUTES = ("008U", "010S", "014U", "058U", "178S", "210U")
EXPECTED_TSMIS_TOTAL = 16_459
EXPECTED_TSN_TOTAL = 16_626
EXPECTED_TSMIS_MINUS_TSN = -167

# Filled only after the independent serializers have completed one exact bound-source
# run.  A placeholder is intentionally fatal; acceptance cannot bless an unfrozen
# digest merely because the values happen to look plausible.
EXPECTED_TSMIS_CROSS_FORMAT_SHA256 = (
    "9c012be4529d358181010dca4c89d0e0e4a759d9c066248feddf0f7149b2f33a")
EXPECTED_TSMIS_AGGREGATE_SHA256 = (
    "0574e4b69729a00e8ce325bca8d515ad8fa1f472599dd13ebfda5503dd3dc7a6")
EXPECTED_COMPARISON_TRUTH_SHA256 = (
    "60459ed21842e53460e10ddc60c66e1cdbab1bf716b76826a5f4128c8b8fc120")

VISUAL_REVIEW = {
    "sample_routes": ["001", "008U", "010S", "058U", "210U", "905"],
    "pages_reviewed": 6,
    "renderer_cross_check": "Poppler plus pypdfium2",
    "all_217_page2_pixel_census": True,
    "near_black_area_threshold": "10 percent",
    "maximum_observed_near_black_fraction": 0.000522,
    "conclusion": (
        "All sampled page-2 tables were legible and the all-217 pixel census was "
        "normal. The original-detail black preview was a viewer artifact, not source data."),
}


class AuditError(ValueError):
    pass


@dataclass(frozen=True)
class FileEntry:
    name: str
    bytes: int
    sha256: str


@dataclass(frozen=True)
class CategorySpec:
    section: str
    source_section: str
    row: int
    raw_label: str
    code: str
    category: str
    pdf_column: str
    sides: str = "both"

    @property
    def display_label(self) -> str:
        if self.code.isdigit():
            return f"{self.code} lane(s)"
        return self.category.split(": ", 1)[1]


def _spec(section: str, source_section: str, row: int, raw: str, code: str,
          label: str, pdf_column: str, sides: str = "both") -> CategorySpec:
    category = (
        f"{section}: {code} lanes" if code.isdigit()
        else f"{section}: {code} - {label}")
    return CategorySpec(
        section, source_section, row, raw, code, category, pdf_column, sides)


SECTIONS: tuple[tuple[str, str, tuple[CategorySpec, ...]], ...] = (
    ("HIGHWAY GROUP", "HIGHWAY GROUP", (
        _spec("HIGHWAY GROUP", "HIGHWAY GROUP", 7, "R-RIGHT IND ALIGN", "R", "RIGHT IND ALIGN", "left"),
        _spec("HIGHWAY GROUP", "HIGHWAY GROUP", 8, "L-LEFT IND ALIGN", "L", "LEFT IND ALIGN", "left"),
        _spec("HIGHWAY GROUP", "HIGHWAY GROUP", 9, "X-UNCONSTRUCTED", "X", "UNCONSTRUCTED", "left"),
        _spec("HIGHWAY GROUP", "HIGHWAY GROUP", 10, "U-UNDIVIDED", "U", "UNDIVIDED", "left"),
        _spec("HIGHWAY GROUP", "HIGHWAY GROUP", 11, "D-DIVIDED", "D", "DIVIDED", "left"),
    )),
    ("RURAL/URBAN/SUBURBAN", "RURAL/URBAN/SUBURBAN", (
        _spec("RURAL/URBAN/SUBURBAN", "RURAL/URBAN/SUBURBAN", 15, "R-RURAL -I INSIDE CITY", "R", "RURAL -I INSIDE CITY", "left"),
        _spec("RURAL/URBAN/SUBURBAN", "RURAL/URBAN/SUBURBAN", 16, "-O OUTSIDE CITY", "R-O", "RURAL -O OUTSIDE CITY", "left"),
        _spec("RURAL/URBAN/SUBURBAN", "RURAL/URBAN/SUBURBAN", 17, "U-URBAN -I INSIDE CITY", "U", "URBAN -I INSIDE CITY", "left"),
        _spec("RURAL/URBAN/SUBURBAN", "RURAL/URBAN/SUBURBAN", 18, "-O OUTSIDE CITY", "U-O", "URBAN -O OUTSIDE CITY", "left"),
        _spec("RURAL/URBAN/SUBURBAN", "RURAL/URBAN/SUBURBAN", 19, "+-INVALID DATA", "+", "INVALID DATA", "left"),
    )),
    ("INTERSECTION TYPE", "INTERSECTION TYPE", (
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 23, "F-FOUR-LEGGED", "F", "FOUR-LEGGED", "left"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 24, "M-MULTI-LEGGED", "M", "MULTI-LEGGED", "left"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 25, "S-OFFSET", "S", "OFFSET", "left"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 26, "T-TEE", "T", "TEE", "left"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 27, "Y-WYE", "Y", "WYE", "left"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 28, "R-ROUNDABOUT", "R", "ROUNDABOUT", "left", "tsmis"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 29, "C-OTHER CIRCULAR INTERSECTION", "C", "OTHER CIRCULAR INTERSECTION", "left", "tsmis"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 30, "P-MIDBLOCK PED CROSSING (AT GRADE)", "P", "MIDBLOCK PED CROSSING (AT GRADE)", "left", "tsmis"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 31, "Z-OTHER", "Z", "OTHER", "left"),
        _spec("INTERSECTION TYPE", "INTERSECTION TYPE", 32, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "left", "tsmis"),
    )),
    ("LIGHTING TYPE", "LIGHTING TYPE", (
        _spec("LIGHTING TYPE", "LIGHTING TYPE", 36, "N-NO LIGHTING", "N", "NO LIGHTING", "left"),
        _spec("LIGHTING TYPE", "LIGHTING TYPE", 37, "Y-LIGHTING", "Y", "LIGHTING", "left"),
        _spec("LIGHTING TYPE", "LIGHTING TYPE", 38, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "left"),
    )),
    ("CONTROL TYPES", "CONTROL TYPES", (
        _spec("CONTROL TYPES", "CONTROL TYPES", 42, "A-NO CONTROL", "A", "NO CONTROL", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 43, "B-STOP SIGNS ON CROSS ST ONLY", "B", "STOP SIGNS ON CROSS ST ONLY", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 44, "C-STOP SIGNS ON MAINLINE ONLY", "C", "STOP SIGNS ON MAINLINE ONLY", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 45, "D-FOUR-WAY STOP SIGNS", "D", "FOUR-WAY STOP SIGNS", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 46, "E-4-WAY FLASHER (RED/CROSS ST)", "E", "4-WAY FLASHER (RED/CROSS ST)", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 47, "F-4-WAY FLASHER (RED/MAINLINE)", "F", "4-WAY FLASHER (RED/MAINLINE)", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 48, "G-4-WAY FLASHER (RED ON ALL)", "G", "4-WAY FLASHER (RED ON ALL)", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 49, "H-YIELD SIGNS (CROSS ST ONLY)", "H", "YIELD SIGNS (CROSS ST ONLY)", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 50, "I-YIELD SIGNS (MAIN LINE ONLY)", "I", "YIELD SIGNS (MAIN LINE ONLY)", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 51, "R-YIELD ALL WAYS (ROUNDABOUT)", "R", "YIELD ALL WAYS (ROUNDABOUT)", "center", "tsmis"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 52, "S-SIGNALIZED", "S", "SIGNALIZED (incl. TSN J-P)", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 53, "O-PEDESTRIAN HYBRID BEACON", "O", "PEDESTRIAN HYBRID BEACON", "center", "tsmis"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 54, "Q-FLASH BEACON", "Q", "FLASH BEACON", "center", "tsmis"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 55, "Z-OTHER", "Z", "OTHER", "center"),
        _spec("CONTROL TYPES", "CONTROL TYPES", 56, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "center"),
    )),
    ("MAINLINE NUM OF LANES", "MAINLINE NUM OF LANES", (
        *tuple(_spec("MAINLINE NUM OF LANES", "MAINLINE NUM OF LANES", 59 + lane,
                     str(lane), str(lane), "", "center") for lane in range(1, 9)),
        _spec("MAINLINE NUM OF LANES", "MAINLINE NUM OF LANES", 68, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "center"),
    )),
    ("MAINLINE MASTARM", "MAINLINE MASTERARM", (
        _spec("MAINLINE MASTARM", "MAINLINE MASTERARM", 72, "Y-YES", "Y", "YES", "right"),
        _spec("MAINLINE MASTARM", "MAINLINE MASTERARM", 73, "N-NO", "N", "NO", "right"),
        _spec("MAINLINE MASTARM", "MAINLINE MASTERARM", 74, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "right"),
    )),
    ("MAINLINE LEFT CHANNELIZATION", "MAINLINE LEFT CHANNELIZATION", (
        _spec("MAINLINE LEFT CHANNELIZATION", "MAINLINE LEFT CHANNELIZATION", 78, "C-CURBED MEDIAN LEFT TURN CHAN", "C", "CURBED MEDIAN LEFT TURN CHAN", "right"),
        _spec("MAINLINE LEFT CHANNELIZATION", "MAINLINE LEFT CHANNELIZATION", 79, "N-NO LEFT TURN CHANNELIZATION", "N", "NO LEFT TURN CHANNELIZATION", "right"),
        _spec("MAINLINE LEFT CHANNELIZATION", "MAINLINE LEFT CHANNELIZATION", 80, "P-PAINTED LEFT TURN CHAN", "P", "PAINTED LEFT TURN CHAN", "right"),
        _spec("MAINLINE LEFT CHANNELIZATION", "MAINLINE LEFT CHANNELIZATION", 81, "R-RAISED BARS LEFT TURN CHAN", "R", "RAISED BARS LEFT TURN CHAN", "right"),
        _spec("MAINLINE LEFT CHANNELIZATION", "MAINLINE LEFT CHANNELIZATION", 82, "Y-CHANNELIZATION - NOT SPECIFIED", "Y", "CHANNELIZATION NOT SPECIFIED", "right", "tsmis"),
        _spec("MAINLINE LEFT CHANNELIZATION", "MAINLINE LEFT CHANNELIZATION", 83, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "right"),
    )),
    ("MAINLINE RIGHT CHANNELIZATION", "MAINLINE RIGHT CHANNELIZATION", (
        _spec("MAINLINE RIGHT CHANNELIZATION", "MAINLINE RIGHT CHANNELIZATION", 87, "Y-FREE RIGHT TURNS", "Y", "FREE RIGHT TURNS", "right"),
        _spec("MAINLINE RIGHT CHANNELIZATION", "MAINLINE RIGHT CHANNELIZATION", 88, "N-NO FREE RIGHT TURNS", "N", "NO FREE RIGHT TURNS", "right"),
        _spec("MAINLINE RIGHT CHANNELIZATION", "MAINLINE RIGHT CHANNELIZATION", 89, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "right"),
    )),
    ("MAINLINE TRAFFIC FLOW", "MAINLINE TRAFFIC FLOW", (
        _spec("MAINLINE TRAFFIC FLOW", "MAINLINE TRAFFIC FLOW", 93, "N-2 WAY - NO LEFT TURNS", "N", "2 WAY - NO LEFT TURNS", "right"),
        _spec("MAINLINE TRAFFIC FLOW", "MAINLINE TRAFFIC FLOW", 94, "P-2 WAY WITH LEFT TURN", "P", "2 WAY WITH LEFT TURN", "right"),
        _spec("MAINLINE TRAFFIC FLOW", "MAINLINE TRAFFIC FLOW", 95, "R-2 WAY - LEFT TURN RESTRICT", "R", "2 WAY - LEFT TURN RESTRICT", "right"),
        _spec("MAINLINE TRAFFIC FLOW", "MAINLINE TRAFFIC FLOW", 96, "W-ONE WAY TRAFFIC", "W", "ONE WAY TRAFFIC", "right"),
        _spec("MAINLINE TRAFFIC FLOW", "MAINLINE TRAFFIC FLOW", 97, "Z-OTHERS", "Z", "OTHERS", "right"),
        _spec("MAINLINE TRAFFIC FLOW", "MAINLINE TRAFFIC FLOW", 98, "+-NO DATA GIVEN", "+", "NO DATA GIVEN", "right"),
    )),
)

CATEGORIES = tuple(cat for _section, _source, cats in SECTIONS for cat in cats)
TSMIS_ORDER = tuple(cat.category for cat in CATEGORIES) + ("Total Intersections",)
TSN_ORDER = tuple(cat.category for cat in CATEGORIES if cat.sides == "both") + (
    "Total Intersections",)
PDF_COLUMNS = {
    column: tuple(cat for cat in CATEGORIES if cat.pdf_column == column)
    for column in ("left", "center", "right")
}
SECTION_CATEGORIES = {
    section: tuple(cat.category for cat in cats) for section, _source, cats in SECTIONS
}

TSN_ROWS: tuple[tuple[str, int], ...] = (
    ("HIGHWAY GROUP: R - RIGHT IND ALIGN", 166),
    ("HIGHWAY GROUP: L - LEFT IND ALIGN", 152),
    ("HIGHWAY GROUP: X - UNCONSTRUCTED", 0),
    ("HIGHWAY GROUP: U - UNDIVIDED", 10_186),
    ("HIGHWAY GROUP: D - DIVIDED", 6_122),
    ("RURAL/URBAN/SUBURBAN: R - RURAL -I INSIDE CITY", 346),
    ("RURAL/URBAN/SUBURBAN: R-O - RURAL -O OUTSIDE CITY", 8_270),
    ("RURAL/URBAN/SUBURBAN: U - URBAN -I INSIDE CITY", 5_500),
    ("RURAL/URBAN/SUBURBAN: U-O - URBAN -O OUTSIDE CITY", 2_510),
    ("RURAL/URBAN/SUBURBAN: + - INVALID DATA", 0),
    ("INTERSECTION TYPE: F - FOUR-LEGGED", 5_244),
    ("INTERSECTION TYPE: M - MULTI-LEGGED", 141),
    ("INTERSECTION TYPE: S - OFFSET", 540),
    ("INTERSECTION TYPE: T - TEE", 9_553),
    ("INTERSECTION TYPE: Y - WYE", 949),
    ("INTERSECTION TYPE: Z - OTHER", 159),
    ("LIGHTING TYPE: N - NO LIGHTING", 8_738),
    ("LIGHTING TYPE: Y - LIGHTING", 7_888),
    ("LIGHTING TYPE: + - NO DATA GIVEN", 0),
    ("CONTROL TYPES: A - NO CONTROL", 1_760),
    ("CONTROL TYPES: B - STOP SIGNS ON CROSS ST ONLY", 11_880),
    ("CONTROL TYPES: C - STOP SIGNS ON MAINLINE ONLY", 98),
    ("CONTROL TYPES: D - FOUR-WAY STOP SIGNS", 78),
    ("CONTROL TYPES: E - 4-WAY FLASHER (RED/CROSS ST)", 30),
    ("CONTROL TYPES: F - 4-WAY FLASHER (RED/MAINLINE)", 7),
    ("CONTROL TYPES: G - 4-WAY FLASHER (RED ON ALL)", 29),
    ("CONTROL TYPES: H - YIELD SIGNS (CROSS ST ONLY)", 22),
    ("CONTROL TYPES: I - YIELD SIGNS (MAIN LINE ONLY)", 2),
    ("CONTROL TYPES: S - SIGNALIZED (incl. TSN J-P)", 2_648),
    ("CONTROL TYPES: Z - OTHER", 32),
    ("CONTROL TYPES: + - NO DATA GIVEN", 0),
    ("MAINLINE NUM OF LANES: 1 lanes", 1),
    ("MAINLINE NUM OF LANES: 2 lanes", 10_374),
    ("MAINLINE NUM OF LANES: 3 lanes", 578),
    ("MAINLINE NUM OF LANES: 4 lanes", 4_465),
    ("MAINLINE NUM OF LANES: 5 lanes", 227),
    ("MAINLINE NUM OF LANES: 6 lanes", 845),
    ("MAINLINE NUM OF LANES: 7 lanes", 22),
    ("MAINLINE NUM OF LANES: 8 lanes", 111),
    ("MAINLINE NUM OF LANES: + - NO DATA GIVEN", 0),
    ("MAINLINE MASTARM: Y - YES", 2_504),
    ("MAINLINE MASTARM: N - NO", 14_122),
    ("MAINLINE MASTARM: + - NO DATA GIVEN", 0),
    ("MAINLINE LEFT CHANNELIZATION: C - CURBED MEDIAN LEFT TURN CHAN", 1_335),
    ("MAINLINE LEFT CHANNELIZATION: N - NO LEFT TURN CHANNELIZATION", 10_347),
    ("MAINLINE LEFT CHANNELIZATION: P - PAINTED LEFT TURN CHAN", 4_897),
    ("MAINLINE LEFT CHANNELIZATION: R - RAISED BARS LEFT TURN CHAN", 17),
    ("MAINLINE LEFT CHANNELIZATION: + - NO DATA GIVEN", 0),
    ("MAINLINE RIGHT CHANNELIZATION: Y - FREE RIGHT TURNS", 2_046),
    ("MAINLINE RIGHT CHANNELIZATION: N - NO FREE RIGHT TURNS", 14_577),
    ("MAINLINE RIGHT CHANNELIZATION: + - NO DATA GIVEN", 0),
    ("MAINLINE TRAFFIC FLOW: N - 2 WAY - NO LEFT TURNS", 646),
    ("MAINLINE TRAFFIC FLOW: P - 2 WAY WITH LEFT TURN", 15_599),
    ("MAINLINE TRAFFIC FLOW: R - 2 WAY - LEFT TURN RESTRICT", 38),
    ("MAINLINE TRAFFIC FLOW: W - ONE WAY TRAFFIC", 326),
    ("MAINLINE TRAFFIC FLOW: Z - OTHERS", 17),
    ("MAINLINE TRAFFIC FLOW: + - NO DATA GIVEN", 0),
    ("Total Intersections", 16_626),
)
TSN_COUNTS = dict(TSN_ROWS)

EXPECTED_TSMIS_AGGREGATE = {
    "HIGHWAY GROUP: R - RIGHT IND ALIGN": 141,
    "HIGHWAY GROUP: L - LEFT IND ALIGN": 108,
    "HIGHWAY GROUP: X - UNCONSTRUCTED": 0,
    "HIGHWAY GROUP: U - UNDIVIDED": 9_556,
    "HIGHWAY GROUP: D - DIVIDED": 5_978,
    "RURAL/URBAN/SUBURBAN: R - RURAL -I INSIDE CITY": 378,
    "RURAL/URBAN/SUBURBAN: R-O - RURAL -O OUTSIDE CITY": 8_124,
    "RURAL/URBAN/SUBURBAN: U - URBAN -I INSIDE CITY": 4_049,
    "RURAL/URBAN/SUBURBAN: U-O - URBAN -O OUTSIDE CITY": 1_288,
    "RURAL/URBAN/SUBURBAN: + - INVALID DATA": 2_620,
    "INTERSECTION TYPE: F - FOUR-LEGGED": 5_268,
    "INTERSECTION TYPE: M - MULTI-LEGGED": 140,
    "INTERSECTION TYPE: S - OFFSET": 528,
    "INTERSECTION TYPE: T - TEE": 9_401,
    "INTERSECTION TYPE: Y - WYE": 925,
    "INTERSECTION TYPE: R - ROUNDABOUT": 42,
    "INTERSECTION TYPE: C - OTHER CIRCULAR INTERSECTION": 0,
    "INTERSECTION TYPE: P - MIDBLOCK PED CROSSING (AT GRADE)": 0,
    "INTERSECTION TYPE: Z - OTHER": 155,
    "INTERSECTION TYPE: + - NO DATA GIVEN": 0,
    "LIGHTING TYPE: N - NO LIGHTING": 8_526,
    "LIGHTING TYPE: Y - LIGHTING": 7_933,
    "LIGHTING TYPE: + - NO DATA GIVEN": 0,
    "CONTROL TYPES: A - NO CONTROL": 1_676,
    "CONTROL TYPES: B - STOP SIGNS ON CROSS ST ONLY": 11_685,
    "CONTROL TYPES: C - STOP SIGNS ON MAINLINE ONLY": 124,
    "CONTROL TYPES: D - FOUR-WAY STOP SIGNS": 84,
    "CONTROL TYPES: E - 4-WAY FLASHER (RED/CROSS ST)": 37,
    "CONTROL TYPES: F - 4-WAY FLASHER (RED/MAINLINE)": 10,
    "CONTROL TYPES: G - 4-WAY FLASHER (RED ON ALL)": 35,
    "CONTROL TYPES: H - YIELD SIGNS (CROSS ST ONLY)": 23,
    "CONTROL TYPES: I - YIELD SIGNS (MAIN LINE ONLY)": 1,
    "CONTROL TYPES: R - YIELD ALL WAYS (ROUNDABOUT)": 42,
    "CONTROL TYPES: S - SIGNALIZED (incl. TSN J-P)": 2_710,
    "CONTROL TYPES: O - PEDESTRIAN HYBRID BEACON": 0,
    "CONTROL TYPES: Q - FLASH BEACON": 0,
    "CONTROL TYPES: Z - OTHER": 32,
    "CONTROL TYPES: + - NO DATA GIVEN": 0,
    "MAINLINE NUM OF LANES: 1 lanes": 1,
    "MAINLINE NUM OF LANES: 2 lanes": 10_202,
    "MAINLINE NUM OF LANES: 3 lanes": 552,
    "MAINLINE NUM OF LANES: 4 lanes": 4_309,
    "MAINLINE NUM OF LANES: 5 lanes": 230,
    "MAINLINE NUM OF LANES: 6 lanes": 832,
    "MAINLINE NUM OF LANES: 7 lanes": 23,
    "MAINLINE NUM OF LANES: 8 lanes": 110,
    "MAINLINE NUM OF LANES: + - NO DATA GIVEN": 200,
    "MAINLINE MASTARM: Y - YES": 2_498,
    "MAINLINE MASTARM: N - NO": 13_769,
    "MAINLINE MASTARM: + - NO DATA GIVEN": 192,
    "MAINLINE LEFT CHANNELIZATION: C - CURBED MEDIAN LEFT TURN CHAN": 1_324,
    "MAINLINE LEFT CHANNELIZATION: N - NO LEFT TURN CHANNELIZATION": 10_104,
    "MAINLINE LEFT CHANNELIZATION: P - PAINTED LEFT TURN CHAN": 4_793,
    "MAINLINE LEFT CHANNELIZATION: R - RAISED BARS LEFT TURN CHAN": 18,
    "MAINLINE LEFT CHANNELIZATION: Y - CHANNELIZATION NOT SPECIFIED": 28,
    "MAINLINE LEFT CHANNELIZATION: + - NO DATA GIVEN": 192,
    "MAINLINE RIGHT CHANNELIZATION: Y - FREE RIGHT TURNS": 2_044,
    "MAINLINE RIGHT CHANNELIZATION: N - NO FREE RIGHT TURNS": 14_219,
    "MAINLINE RIGHT CHANNELIZATION: + - NO DATA GIVEN": 196,
    "MAINLINE TRAFFIC FLOW: N - 2 WAY - NO LEFT TURNS": 620,
    "MAINLINE TRAFFIC FLOW: P - 2 WAY WITH LEFT TURN": 15_283,
    "MAINLINE TRAFFIC FLOW: R - 2 WAY - LEFT TURN RESTRICT": 36,
    "MAINLINE TRAFFIC FLOW: W - ONE WAY TRAFFIC": 310,
    "MAINLINE TRAFFIC FLOW: Z - OTHERS": 18,
    "MAINLINE TRAFFIC FLOW: + - NO DATA GIVEN": 192,
    "Total Intersections": 16_459,
}

EXPECTED_IDENTICAL = (
    "HIGHWAY GROUP: X - UNCONSTRUCTED",
    "LIGHTING TYPE: + - NO DATA GIVEN",
    "CONTROL TYPES: Z - OTHER",
    "CONTROL TYPES: + - NO DATA GIVEN",
    "MAINLINE NUM OF LANES: 1 lanes",
)
EXPECTED_ONLY_TSMIS = (
    "INTERSECTION TYPE: R - ROUNDABOUT",
    "INTERSECTION TYPE: C - OTHER CIRCULAR INTERSECTION",
    "INTERSECTION TYPE: P - MIDBLOCK PED CROSSING (AT GRADE)",
    "INTERSECTION TYPE: + - NO DATA GIVEN",
    "CONTROL TYPES: R - YIELD ALL WAYS (ROUNDABOUT)",
    "CONTROL TYPES: O - PEDESTRIAN HYBRID BEACON",
    "CONTROL TYPES: Q - FLASH BEACON",
    "MAINLINE LEFT CHANNELIZATION: Y - CHANNELIZATION NOT SPECIFIED",
)


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_digest(rows: Iterable[Sequence[object]]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update((json.dumps(
            list(row), ensure_ascii=False, separators=(",", ":")) + "\n").encode("utf-8"))
    return digest.hexdigest()


def _file_identity(path: Path) -> dict[str, object]:
    return {"bytes": path.stat().st_size, "sha256": _sha_file(path)}


def _strict_count(value: object, context: str) -> int:
    if type(value) is not int or value < 0:
        raise AuditError(f"{context}: expected nonnegative integer, got {value!r}")
    return value


def _route_from_name(path: Path, suffix: str) -> str:
    match = re.fullmatch(
        rf"intersection_summary_route_(\d{{3}}[A-Z]?)\{suffix}", path.name)
    if not match:
        raise AuditError(f"unexpected Intersection Summary filename: {path.name}")
    return match.group(1)


def _manifest(root: Path, suffix: str) -> tuple[dict[str, object], list[FileEntry]]:
    paths = sorted(root.glob(f"*{suffix}"), key=lambda path: path.name)
    entries = [FileEntry(
        path.name, path.stat().st_size, _sha_file(path).lower()) for path in paths]
    serialized = "".join(
        f"{entry.name}\t{entry.bytes}\t{entry.sha256}\n" for entry in entries
    ).encode("utf-8")
    return {
        "files": len(entries),
        "bytes": sum(entry.bytes for entry in entries),
        "manifest_sha256": _sha_bytes(serialized),
        "serialization": "name\\tbytes\\tlowercase-member-sha256\\n sorted by name",
    }, entries


def _capture_tree(label: str, root: Path, binding: dict[str, object],
                  destination: Path) -> tuple[dict[str, object], Path]:
    observed, entries = _manifest(root, str(binding["suffix"]))
    for key in ("files", "bytes", "manifest_sha256"):
        if observed[key] != binding[key]:
            raise AuditError(
                f"{label} {key} drift: {observed[key]!r} != {binding[key]!r}")
    destination.mkdir(parents=True, exist_ok=False)
    for entry in entries:
        payload = (root / entry.name).read_bytes()
        if len(payload) != entry.bytes or _sha_bytes(payload) != entry.sha256:
            raise AuditError(f"{label} changed during capture: {entry.name}")
        (destination / entry.name).write_bytes(payload)
    captured, _ = _manifest(destination, str(binding["suffix"]))
    if captured != observed:
        raise AuditError(f"{label} private snapshot differs from source")
    return {
        "binding": dict(binding), "observed": observed,
        "members": [asdict(entry) for entry in entries],
    }, destination


def _capture_file(label: str, source: Path, binding: dict[str, object],
                  destination: Path) -> tuple[dict[str, object], Path]:
    payload = source.read_bytes()
    observed = {"bytes": len(payload), "sha256": _sha_bytes(payload)}
    if observed != binding:
        raise AuditError(f"{label} identity drift: {observed!r} != {binding!r}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(payload)
    if destination.read_bytes() != payload:
        raise AuditError(f"{label} private snapshot changed after write")
    return {"binding": dict(binding), "observed": observed}, destination


def _route_lf_digest(routes: Sequence[str]) -> str:
    return _sha_bytes("".join(f"{route}\n" for route in routes).encode("utf-8"))


def _require_exact_route_universes(
        universes: dict[str, Sequence[str]], expected_count: int = EXPECTED_ROUTE_COUNT,
        expected_digest: str | None = EXPECTED_ROUTE_LF_SHA256) -> dict[str, object]:
    if not universes:
        raise AuditError("no route universes supplied")
    canonical_label, canonical_values = next(iter(universes.items()))
    canonical = list(canonical_values)
    detail = {}
    for label, raw_values in universes.items():
        values = list(raw_values)
        duplicates = sorted(
            route for route, count in Counter(values).items() if count != 1)
        if len(values) != expected_count or duplicates:
            raise AuditError(
                f"{label} route universe invalid: rows={len(values)} "
                f"duplicates={duplicates!r}")
        if values != canonical:
            raise AuditError(
                f"{label} route universe/order differs from {canonical_label}: "
                f"missing={sorted(set(canonical)-set(values))!r} "
                f"extra={sorted(set(values)-set(canonical))!r}")
        digest = _route_lf_digest(values)
        if expected_digest is not None and digest != expected_digest:
            raise AuditError(
                f"{label} route digest drift: {digest} != {expected_digest}")
        detail[label] = {"routes": len(values), "unique": len(set(values)),
                         "lf_sha256": digest}
    suffixes = tuple(route for route in canonical if not route.isdigit())
    if expected_count == EXPECTED_ROUTE_COUNT:
        if suffixes != EXPECTED_SUFFIX_ROUTES or "170" in canonical:
            raise AuditError(
                f"authoritative suffix/170 contract drift: {suffixes!r}")
    return {
        "canonical": canonical_label, "expected_count": expected_count,
        "routes": canonical, "suffix_routes": list(suffixes),
        "route_170_absent": "170" not in canonical,
        "universes": detail, "all_exact": True,
    }


def _expected_xlsx_static_cells(route: str, total: int) -> dict[tuple[int, int], object]:
    expected: dict[tuple[int, int], object] = {
        (1, 1): "TSAR - Intersection Summary",
        (2, 1): f"Route: {route}",
        (3, 1): f"Total Intersections = {total}",
    }
    for _section, source_section, cats in SECTIONS:
        heading_row = cats[0].row - 2
        expected[(heading_row, 1)] = source_section
        expected[(heading_row + 1, 1)] = "NUMBER"
        expected[(heading_row + 1, 2)] = (
            "LANES" if source_section == "MAINLINE NUM OF LANES" else "CODE")
        for cat in cats:
            expected[(cat.row, 2)] = cat.raw_label
    return expected


def _validate_record(record: dict[str, int], context: str) -> dict[str, int]:
    if tuple(record) != TSMIS_ORDER:
        raise AuditError(f"{context}: category order/universe drift")
    typed = {key: _strict_count(value, f"{context} {key}")
             for key, value in record.items()}
    total = typed["Total Intersections"]
    if total <= 0:
        raise AuditError(f"{context}: Total Intersections must be positive")
    section_sums = {
        section: sum(typed[key] for key in keys)
        for section, keys in SECTION_CATEGORIES.items()
    }
    for section, subtotal in section_sums.items():
        if section == "HIGHWAY GROUP":
            continue
        if subtotal != total:
            raise AuditError(
                f"{context}: {section} subtotal {subtotal} != total {total}")
    return section_sums


def _parse_tsmis_xlsx(root: Path) -> dict[str, object]:
    records: dict[str, dict[str, int]] = {}
    section_rows = []
    for path in sorted(root.glob("*.xlsx"), key=lambda item: item.name):
        route = _route_from_name(path, ".xlsx")
        if route in records:
            raise AuditError(f"duplicate XLSX route {route}")
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            if workbook.sheetnames != ["Intersection Summary"]:
                raise AuditError(f"{path.name}: sheet universe drift")
            ws = workbook["Intersection Summary"]
            if (ws.max_row, ws.max_column) != (99, 3):
                raise AuditError(
                    f"{path.name}: expected 99x3, got {ws.max_row}x{ws.max_column}")
            route_cell = ws["A2"].value
            if route_cell != f"Route: {route}":
                raise AuditError(f"{path.name}: internal route drift {route_cell!r}")
            total_text = ws["A3"].value
            match = re.fullmatch(r"Total Intersections = ([\d,]+)", str(total_text))
            if not match:
                raise AuditError(f"{path.name}: malformed/missing total")
            total = int(match.group(1).replace(",", ""))
            static = _expected_xlsx_static_cells(route, total)
            data_rows = {cat.row: cat for cat in CATEGORIES}
            record: dict[str, int] = {}
            for row in range(1, 100):
                for column in range(1, 4):
                    observed = ws.cell(row, column).value
                    if row in data_rows and column == 1:
                        value = _strict_count(
                            observed, f"{path.name} row {row} count")
                        record[data_rows[row].category] = value
                    else:
                        expected = static.get((row, column))
                        if observed != expected:
                            raise AuditError(
                                f"{path.name} {ws.cell(row,column).coordinate}: "
                                f"{observed!r} != {expected!r}")
            record["Total Intersections"] = total
            section_sums = _validate_record(record, path.name)
            section_rows.extend(
                (route, section, subtotal, total)
                for section, subtotal in section_sums.items())
            records[route] = record
        finally:
            workbook.close()
    return {
        "routes": list(records), "records": records,
        "files": len(records), "sheets": len(records),
        "source_value_cells": len(records) * len(TSMIS_ORDER),
        "section_arithmetic_sha256": _canonical_digest(section_rows),
        "all_non_highway_partitions_exact": True,
    }


def _validate_pdf_provenance(page1_text: str, page2_text: str,
                             metadata: dict[str, str], route: str,
                             filename: str) -> dict[str, object]:
    expected_metadata = {
        "Title": "TSMIS Reports",
        "Creator": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) HeadlessChrome/150.0.0.0 Safari/537.36"),
        "Producer": "Skia/PDF m150",
    }
    if any(metadata.get(key) != value for key, value in expected_metadata.items()):
        raise AuditError(f"{filename}: PDF metadata producer/title drift")
    creation = metadata.get("CreationDate")
    if (creation != metadata.get("ModDate")
            or not re.fullmatch(r"D:20260710\d{6}\+00'00'", str(creation))):
        raise AuditError(f"{filename}: PDF creation/modification metadata drift")
    required_page1 = (
        "REPORT DATE : 07/09/2026",
        "REFERENCE DATE : 07/10/2026",
        "SUBMITTOR : Yunus.Shaikh@dot.ca.gov",
        "REPORT TITLE : TSAR - Intersection Summary",
        f"ROUTE {route}",
        "PM ALL",
    )
    if any(token not in page1_text for token in required_page1):
        raise AuditError(f"{filename}: page-1 provenance/route drift")
    required_page2 = (
        "TSAR – Intersection Summary",
        "District: ALL County: ALL",
        f"Route: {route}",
        "Ref Date: 2026-07-10",
        "Total Intersections =",
    )
    if any(token not in page2_text for token in required_page2):
        raise AuditError(f"{filename}: page-2 header/provenance drift")
    normalized = re.sub(r"[^A-Z]+", " ", page2_text.upper())
    headings = tuple(source for _section, source, _cats in SECTIONS)
    if any(re.sub(r"[^A-Z]+", " ", heading).strip() not in normalized
           for heading in headings):
        raise AuditError(f"{filename}: page-2 section heading drift")
    return {
        "report_date": "07/09/2026", "reference_date": "07/10/2026",
        "submitter": "Yunus.Shaikh@dot.ca.gov", "route": route,
        "creation": creation,
    }


def _pdf_count_words(words: Sequence[dict[str, object]], filename: str
                     ) -> dict[str, list[dict[str, object]]]:
    selected = {"left": [], "center": [], "right": []}
    for word in words:
        text = str(word["text"])
        if not re.fullmatch(r"[\d,]+", text):
            continue
        x1 = float(word["x1"])
        top = float(word["top"])
        column = None
        if 50 <= x1 <= 70 and 120 <= top <= 540:
            column = "left"
        elif 230 <= x1 <= 260 and 120 <= top <= 480:
            column = "center"
        elif 415 <= x1 <= 450 and 120 <= top <= 480:
            column = "right"
        if column is not None:
            selected[column].append(word)
    for column in selected:
        selected[column].sort(key=lambda word: (float(word["top"]), float(word["x0"])))
        expected = len(PDF_COLUMNS[column])
        if len(selected[column]) != expected:
            raise AuditError(
                f"{filename}: PDF {column} count geometry {len(selected[column])} != {expected}")
    return selected


def _parse_tsmis_pdfs(root: Path) -> dict[str, object]:
    records: dict[str, dict[str, int]] = {}
    metadata_signatures: Counter[tuple[tuple[str, str], ...]] = Counter()
    geometry_signatures: Counter[str] = Counter()
    provenance_rows = []
    for path in sorted(root.glob("*.pdf"), key=lambda item: item.name):
        route = _route_from_name(path, ".pdf")
        if route in records:
            raise AuditError(f"duplicate PDF route {route}")
        with pdfplumber.open(path) as pdf:
            if len(pdf.pages) != 2:
                raise AuditError(f"{path.name}: expected 2 pages, got {len(pdf.pages)}")
            if any((float(page.width), float(page.height)) != (612.0, 792.0)
                   for page in pdf.pages):
                raise AuditError(f"{path.name}: page geometry drift")
            metadata = {str(key): str(value)
                        for key, value in (pdf.metadata or {}).items()}
            page1_text = pdf.pages[0].extract_text(x_tolerance=2, y_tolerance=2) or ""
            page2_text = pdf.pages[1].extract_text(x_tolerance=2, y_tolerance=2) or ""
            provenance = _validate_pdf_provenance(
                page1_text, page2_text, metadata, route, path.name)
            total_match = re.search(
                r"Total Intersections\s*=\s*([\d,]+)", page2_text)
            if not total_match:
                raise AuditError(f"{path.name}: missing page-2 total")
            total = int(total_match.group(1).replace(",", ""))
            selected = _pdf_count_words(pdf.pages[1].extract_words(), path.name)
            record: dict[str, int] = {}
            geometry_rows = []
            for column in ("left", "center", "right"):
                for cat, word in zip(PDF_COLUMNS[column], selected[column], strict=True):
                    value = int(str(word["text"]).replace(",", ""))
                    record[cat.category] = _strict_count(
                        value, f"{path.name} {cat.category}")
                    geometry_rows.append((column, cat.category, round(float(word["top"]), 2)))
            # PDF columns are physical, not canonical section order.
            record = {cat.category: record[cat.category] for cat in CATEGORIES}
            record["Total Intersections"] = total
            _validate_record(record, path.name)
            records[route] = record
            metadata_signatures[tuple(sorted(metadata.items()))] += 1
            geometry_signatures[_canonical_digest(geometry_rows)] += 1
            provenance_rows.append((
                route, provenance["report_date"], provenance["reference_date"],
                provenance["submitter"], provenance["creation"]))
    if len(metadata_signatures) != 173:
        raise AuditError(
            f"TSMIS PDF metadata signature count {len(metadata_signatures)} != 173")
    if len(geometry_signatures) != 1:
        raise AuditError(
            f"TSMIS PDF numeric geometry signatures differ: {len(geometry_signatures)}")
    return {
        "routes": list(records), "records": records, "files": len(records),
        "pages": len(records) * 2,
        "source_value_cells": len(records) * len(TSMIS_ORDER),
        "metadata_signature_count": len(metadata_signatures),
        "metadata_signatures": [
            {"members": count, "metadata": dict(signature)}
            for signature, count in sorted(metadata_signatures.items())],
        "numeric_geometry_sha256": next(iter(geometry_signatures)),
        "provenance_sha256": _canonical_digest(provenance_rows),
    }


def _cross_format(xlsx_records: dict[str, dict[str, int]],
                  pdf_records: dict[str, dict[str, int]]) -> dict[str, object]:
    rows = []
    differences = []
    for route in xlsx_records:
        if route not in pdf_records:
            differences.append({"route": route, "reason": "missing_pdf"})
            continue
        for category in TSMIS_ORDER:
            left = xlsx_records[route][category]
            right = pdf_records[route][category]
            rows.append((route, category, left, right))
            if left != right:
                differences.append({
                    "route": route, "category": category,
                    "xlsx": left, "pdf": right})
    for route in sorted(set(pdf_records) - set(xlsx_records)):
        differences.append({"route": route, "reason": "missing_xlsx"})
    return {
        "routes": len(xlsx_records), "categories_including_total": len(TSMIS_ORDER),
        "paired_values": len(rows), "difference_count": len(differences),
        "differences": differences,
        "ordered_typed_sha256": _canonical_digest(rows),
        "all_exact": not differences,
    }


def _aggregate(records: dict[str, dict[str, int]]) -> dict[str, int]:
    return {key: sum(record[key] for record in records.values())
            for key in TSMIS_ORDER}


def _load_tsn_normalized(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != ["Intersection Summary (TSN)"]:
            raise AuditError("TSN normalized sheet universe drift")
        ws = workbook["Intersection Summary (TSN)"]
        if (ws.max_row, ws.max_column) != (59, 2):
            raise AuditError(
                f"TSN normalized dimensions {ws.max_row}x{ws.max_column} != 59x2")
        if (ws["A1"].value, ws["B1"].value) != ("Category", "Count"):
            raise AuditError("TSN normalized header drift")
        rows = []
        seen = set()
        for row_number in range(2, 60):
            category = ws.cell(row_number, 1).value
            count = _strict_count(
                ws.cell(row_number, 2).value, f"TSN row {row_number} Count")
            if not isinstance(category, str) or not category:
                raise AuditError(f"TSN row {row_number}: invalid category {category!r}")
            if category in seen:
                raise AuditError(f"TSN normalized duplicate category {category!r}")
            seen.add(category)
            rows.append((category, count))
        if tuple(rows) != TSN_ROWS:
            raise AuditError("TSN normalized category order/count truth drift")
        return {
            "sheet": ws.title, "rows": len(rows), "order": [row[0] for row in rows],
            "counts": dict(rows), "ordered_typed_sha256": _canonical_digest(rows),
        }
    finally:
        workbook.close()


def _load_tsn_raw_pdf(path: Path) -> dict[str, object]:
    with pdfplumber.open(path) as pdf:
        if len(pdf.pages) != 3:
            raise AuditError(f"TSN raw PDF pages {len(pdf.pages)} != 3")
        metadata = {str(key): str(value)
                    for key, value in (pdf.metadata or {}).items()}
        expected_metadata = {
            "Creator": "Oracle12c AS Reports Services",
            "CreationDate": "D:20250915165412",
            "ModDate": "D:20250915165412",
            "Producer": "Oracle PDF driver",
            "Title": "otm22250.pdf",
            "Author": "Oracle Reports",
        }
        if metadata != expected_metadata:
            raise AuditError("TSN raw PDF metadata drift")
        texts = [page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                 for page in pdf.pages]
        required = (
            "OTM22250", "REPORT DATE : 09/15/2025",
            "REFERENCE DATE : 09/15/2025", "SUBMITTOR : TRLBUGNI",
            "EVENT ID : 4843738", "STATEWIDE", "04:53 PM",
            "Total Intersections = 16626",
            "F-FOUR WAY FLASHER (RED ON ALL)",
            "G-FOUR WAY FLASHER (RED ON ALL)",
        )
        joined = "\n".join(texts)
        if any(token not in joined for token in required):
            raise AuditError("TSN raw PDF provenance/Control-F text drift")
        return {
            "pages": 3, "metadata": metadata,
            "page_text_sha256": [_sha_bytes(text.encode("utf-8")) for text in texts],
            "report_provenance": {
                "report_id": "OTM22250", "report_date": "09/15/2025",
                "reference_date": "09/15/2025", "submitter": "TRLBUGNI",
                "report_title": "Intersection Summary Statewide",
                "event_id": "4843738", "location": "STATEWIDE",
                "printed_generation_time": "04:53 PM",
                "total_intersections": 16_626,
            },
            "raw_control_f_label": "F-FOUR WAY FLASHER (RED ON ALL)",
            "raw_control_g_label": "G-FOUR WAY FLASHER (RED ON ALL)",
        }


def _load_stage6(result_path: Path, acceptance_path: Path) -> dict[str, object]:
    result = json.loads(result_path.read_text(encoding="utf-8"))
    acceptance = json.loads(acceptance_path.read_text(encoding="utf-8"))
    if acceptance.get("result_sha256") != _sha_file(result_path):
        raise AuditError("Stage-6 acceptance does not bind the exact result")
    invariants = result.get("audit_invariants")
    if not isinstance(invariants, dict) or len(invariants) != 19 or not all(
            value is True for value in invariants.values()):
        raise AuditError("Stage-6 invariant set is not exact/all true")
    projection = result.get("projection_comparison", {})
    fold = projection.get("legacy_signal_fold", {})
    control_f = projection.get("control_f_label_classification", {})
    expected_fold = {"J": 207, "K": 36, "L": 107,
                     "M": 65, "N": 210, "P": 2023}
    if (result.get("projection_exact") is not True
            or result.get("stage6_family_audit_complete") is not True
            or result.get("normalized_full_conservation") is not False
            or projection.get("ordered_exact") is not True
            or projection.get("multiset_exact") is not True
            or projection.get("unexplained_residue") != []
            or fold.get("source_counts") != expected_fold
            or fold.get("folded_count") != 2_648
            or control_f.get("source_descriptor")
            != "F-FOUR WAY FLASHER (RED ON ALL)"
            or control_f.get("normalized_category")
            != "CONTROL TYPES: F - 4-WAY FLASHER (RED/MAINLINE)"):
        raise AuditError("Stage-6 accepted Intersection Summary claims drift")
    if (len(result.get("source_category_dispositions", [])) != 62
            or len(result.get("per_normalized_category_conservation", [])) != 58):
        raise AuditError("Stage-6 source/target ledger cardinality drift")
    return {
        "result_sha256": _sha_file(result_path),
        "acceptance_sha256": _sha_file(acceptance_path),
        "acceptance_result_sha256": acceptance["result_sha256"],
        "projection_exact": True, "stage6_family_audit_complete": True,
        "normalized_full_conservation": False,
        "audit_invariant_count": len(invariants),
        "raw_source_rows": 62, "normalized_rows": 58,
        "signal_fold": fold, "control_f_source_classification": control_f,
    }


def _load_cross_format_result(path: Path) -> dict[str, object]:
    result = json.loads(path.read_text(encoding="utf-8"))
    detail = result.get("detail_cross_format", {})
    summary = result.get("summary_from_detail", {})
    mapping = result.get("report_view_source_mapping", {})
    source_contract = result.get("source_contract", {})
    if (result.get("status") != "pass" or result.get("unresolved_gate_count") != 0
            or source_contract.get("pass") is not True
            or detail.get("unresolved") != []
            or summary.get("unresolved") != []
            or summary.get("unresolved_mismatch_members") != []
            or summary.get("raw_category_delta_count") != 0
            or summary.get("normalized_delta_count") != 0
            or mapping.get("all_xlsx_columns_mapped_once") is not True):
        raise AuditError("accepted TSN Detail/PDF/Summary cross-format oracle drift")
    sources = result.get("sources", {})
    summary_pdf = sources.get("summary_pdf", {})
    if (summary_pdf.get("sha256") != FILE_BINDINGS["tsn_raw_pdf"]["sha256"]
            or summary_pdf.get("size") != FILE_BINDINGS["tsn_raw_pdf"]["bytes"]):
        raise AuditError("cross-format oracle binds a different TSN Summary PDF")
    return {
        "status": "pass", "unresolved_gate_count": 0,
        "detail_xlsx_records": detail.get("xlsx_records"),
        "detail_pdf_records": detail.get("pdf_records"),
        "detail_asserted_cells": detail.get("asserted_cells"),
        "summary_raw_rows": summary.get("raw_pdf_category_count"),
        "summary_normalized_rows_including_total": summary.get(
            "normalized_categories_including_total"),
        "all_report_view_columns_mapped_once": True,
    }


def _load_tsnr_reference(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise AuditError("TSNR reference sheet universe drift")
        ws = workbook["Sheet1"]
        if (ws.max_row, ws.max_column) != (3, 8):
            raise AuditError("TSNR reference dimensions drift")
        current = str(ws["B3"].value or "")
        target = str(ws["E3"].value or "")
        mapping = str(ws["H3"].value or "")
        required_current = (
            "F. Four-Way Flasher (Red on Mainline)",
            "G. Four-Way flasher (Red on All)",
            "J. Signals Pretimed", "K. Signals Pretimed",
            "L. Signals Semi-Traffic Actuated",
            "M. Signals Semi-Traffic Actuated",
            "N. Signals Full-Traffic Actuated",
            "P. Signals Full-Traffic Actuated",
        )
        required_mapping = tuple(f"TSN: {code}" for code in "JKLMNP")
        if (any(token not in current for token in required_current)
                or "5. Signalized" not in target
                or any(token not in mapping for token in required_mapping)
                or mapping.count("Signalized") < 6):
            raise AuditError("TSNR F/G or J-P mapping contract drift")
        return {
            "sheet": ws.title, "shape": [ws.max_row, ws.max_column],
            "current_control_sha256": _sha_bytes(current.encode("utf-8")),
            "target_control_sha256": _sha_bytes(target.encode("utf-8")),
            "mapping_sha256": _sha_bytes(mapping.encode("utf-8")),
            "control_f_canonical": "Four-Way Flasher (Red on Mainline)",
            "control_g_canonical": "Four-Way flasher (Red on All)",
            "jp_fold_target": "5. Signalized", "jp_members": list("JKLMNP"),
            "semantic_mapping_proven": True,
        }
    finally:
        workbook.close()


def _comparison_truth(tsmis: dict[str, int], tsn: dict[str, int]
                      ) -> dict[str, object]:
    rows = []
    for category in TSMIS_ORDER:
        left = tsmis.get(category)
        right = tsn.get(category)
        status = (
            "Both" if left is not None and right is not None
            else "TSMIS only" if left is not None else "TSN only")
        differing = bool(left != right) if status == "Both" else None
        rows.append({
            "category": category, "status": status,
            "tsmis": left, "tsn": right, "differing": differing,
        })
    counts = {
        "union_rows": len(rows),
        "paired_rows": sum(row["status"] == "Both" for row in rows),
        "tsmis_only_rows": sum(row["status"] == "TSMIS only" for row in rows),
        "tsn_only_rows": sum(row["status"] == "TSN only" for row in rows),
        "differing_shared_rows": sum(row["differing"] is True for row in rows),
        "identical_shared_rows": sum(row["differing"] is False for row in rows),
    }
    expected_counts = {
        "union_rows": 66, "paired_rows": 58, "tsmis_only_rows": 8,
        "tsn_only_rows": 0, "differing_shared_rows": 53,
        "identical_shared_rows": 5,
    }
    if counts != expected_counts:
        raise AuditError(f"comparison truth count drift: {counts!r}")
    if tuple(row["category"] for row in rows if row["status"] == "TSMIS only") \
            != EXPECTED_ONLY_TSMIS:
        raise AuditError("comparison TSMIS-only category universe drift")
    if tuple(row["category"] for row in rows if row["differing"] is False) \
            != EXPECTED_IDENTICAL:
        raise AuditError("comparison identical category set drift")
    digest = _canonical_digest((
        row["category"], row["status"], row["tsmis"], row["tsn"], row["differing"])
        for row in rows)
    return {"rows": rows, "counts": counts, "ordered_typed_sha256": digest}


def _typed_cell_value(value: object) -> dict[str, object]:
    if value is None:
        return {"type": "none", "value": None}
    if isinstance(value, bool):
        return {"type": "bool", "value": value}
    if isinstance(value, int):
        return {"type": "int", "value": value}
    if isinstance(value, float):
        return {"type": "float", "value": repr(value)}
    if isinstance(value, (datetime, date, time)):
        return {"type": type(value).__name__, "value": value.isoformat()}
    if isinstance(value, bytes):
        return {"type": "bytes", "value": value.hex()}
    return {"type": type(value).__name__, "value": str(value)}


def _zip_digest_without_core(path: Path) -> dict[str, object]:
    rows = []
    excluded = []
    with zipfile.ZipFile(path) as archive:
        names = [info.filename for info in archive.infolist()]
        duplicates = sorted(
            name for name, count in Counter(names).items() if count != 1)
        if duplicates:
            raise AuditError(f"{path.name}: duplicate ZIP members {duplicates!r}")
        bad = archive.testzip()
        if bad is not None:
            raise AuditError(f"{path.name}: corrupt ZIP member {bad!r}")
        for name in sorted(names):
            payload = archive.read(name)
            if name == "docProps/core.xml":
                excluded.append(name)
                continue
            rows.append((name, len(payload), _sha_bytes(payload)))
    return {
        "member_count": len(names), "included_member_count": len(rows),
        "included_uncompressed_bytes": sum(row[1] for row in rows),
        "excluded_members": excluded,
        "canonical_member_sha256": _canonical_digest(rows),
    }


def _workbook_semantic_digest(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    rows: list[tuple[object, ...]] = []
    cell_count = 0
    formula_count = 0
    try:
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            ws = workbook[sheet_name]
            rows.append((
                "sheet", sheet_index, sheet_name, ws.sheet_state,
                ws.max_row, ws.max_column, str(ws.freeze_panes or ""),
                str(ws.auto_filter.ref or "")))
            for merged in sorted(str(item) for item in ws.merged_cells.ranges):
                rows.append(("merge", sheet_name, merged))
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    typed = _typed_cell_value(cell.value)
                    rows.append((
                        "cell", sheet_name, cell.coordinate, cell.data_type,
                        typed["type"], typed["value"], cell.number_format))
                    cell_count += 1
                    formula_count += cell.data_type == "f"
    finally:
        workbook.close()
    return {
        "sheets": len([row for row in rows if row[0] == "sheet"]),
        "nonblank_cells": cell_count, "formula_cells": formula_count,
        "ordered_semantic_sha256": _canonical_digest(rows),
    }


def _product_int(value: object, context: str) -> int:
    if isinstance(value, bool):
        raise AuditError(f"{context}: Boolean is not a product count")
    if isinstance(value, int) and value >= 0:
        return value
    if isinstance(value, str) and re.fullmatch(r"0|[1-9]\d*", value):
        return int(value)
    raise AuditError(f"{context}: expected integer literal, got {value!r}")


def _workbook_strings(workbook: openpyxl.Workbook) -> list[str]:
    return [
        cell.value
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]


def _inspect_production_consolidated(
        path: Path, raw_records: dict[str, dict[str, int]],
        aggregate: dict[str, int]) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    mismatches = []
    try:
        if workbook.sheetnames != ["Combined", "Intersection Summary"]:
            raise AuditError(
                f"product consolidated sheet universe {workbook.sheetnames!r}")
        ws = workbook["Intersection Summary"]
        if (ws.max_row, ws.max_column) != (218, 67):
            raise AuditError(
                f"product consolidated dimensions {ws.max_row}x{ws.max_column}")
        expected_headers = ("Route", "Total Intersections", *TSMIS_ORDER[:-1])
        observed_headers = tuple(ws.cell(1, column).value for column in range(1, 68))
        if observed_headers != expected_headers:
            raise AuditError("product consolidated header order/universe drift")
        found_routes = []
        for row_number, route in enumerate(raw_records, 2):
            observed_route = ws.cell(row_number, 1).value
            if observed_route != route or not isinstance(observed_route, str):
                raise AuditError(
                    f"product consolidated route row {row_number}: {observed_route!r}")
            found_routes.append(route)
            expected_values = (
                raw_records[route]["Total Intersections"],
                *(raw_records[route][key] for key in TSMIS_ORDER[:-1]))
            for offset, expected in enumerate(expected_values, 2):
                cell = ws.cell(row_number, offset)
                if type(cell.value) is not int or cell.value != expected \
                        or cell.data_type == "f":
                    mismatches.append({
                        "route": route, "cell": cell.coordinate,
                        "expected": expected, "observed": cell.value,
                        "type": type(cell.value).__name__,
                    })
        if found_routes != list(raw_records):
            raise AuditError("product consolidated route order/universe drift")

        combined = workbook["Combined"]
        if (combined.max_row, combined.max_column) != (87, 2):
            raise AuditError(
                f"product Combined dimensions {combined.max_row}x{combined.max_column}")
        if (combined["A1"].value, combined["A2"].value, combined["B2"].value) != (
                "All Routes Combined — TSAR Intersection Summary",
                "Total Intersections", EXPECTED_TSMIS_TOTAL):
            raise AuditError("product Combined title/total drift")
        row_number = 4
        combined_rows = []
        for section, _source, cats in SECTIONS:
            if (combined.cell(row_number, 1).value,
                    combined.cell(row_number, 2).value) != (section, None):
                raise AuditError(
                    f"product Combined section row {row_number} drift")
            row_number += 1
            for cat in cats:
                observed = (
                    combined.cell(row_number, 1).value,
                    combined.cell(row_number, 2).value)
                expected = (cat.display_label, aggregate[cat.category])
                if observed != expected or type(observed[1]) is not int:
                    mismatches.append({
                        "combined_row": row_number,
                        "expected": expected, "observed": observed})
                combined_rows.append((section, cat.display_label, observed[1]))
                row_number += 1
            row_number += 1
        if row_number != 89:
            raise AuditError(f"product Combined row cursor drift: {row_number}")
        strings = _workbook_strings(workbook)
        provenance_tokens = (
            "07/09/2026", "07/10/2026", "Yunus.Shaikh@dot.ca.gov",
            "REPORT DATE", "REFERENCE DATE", "SUBMITTOR", "CreationDate",
        )
        provenance_presence = {
            token: any(token in value for value in strings)
            for token in provenance_tokens
        }
    finally:
        workbook.close()
    if mismatches:
        raise AuditError(
            f"product consolidated value/view drift: {mismatches[:3]!r}")
    return {
        "package_without_core": _zip_digest_without_core(path),
        "semantic_digest": _workbook_semantic_digest(path),
        "routes": len(raw_records),
        "source_backed_values_compared": len(raw_records) * len(TSMIS_ORDER),
        "source_backed_value_mismatches": 0,
        "combined_category_rows": len(CATEGORIES),
        "combined_ordered_typed_sha256": _canonical_digest(combined_rows),
        "printed_tsmis_provenance_presence": provenance_presence,
        "printed_tsmis_provenance_preserved": any(provenance_presence.values()),
        "projection_exact": True,
    }


def _expected_comparison_formulas(row_number: int, ordinal: int) -> tuple[str, ...]:
    key = f"__CMP_E2_KEY_V1_{ordinal:08d}"
    link = []
    for sheet in ("TSMIS", "TSN"):
        link.append(
            f'=IFERROR(HYPERLINK("#{sheet}!"&MATCH("{key}",{sheet}!$D:$D,0)&":"&'
            f'MATCH("{key}",{sheet}!$D:$D,0),MATCH("{key}",{sheet}!$D:$D,0)),"")')
    status = (
        f'=IF(AND(C{row_number}<>"",D{row_number}<>""),"Both",'
        f'IF(C{row_number}<>"","TSMIS only","TSN only"))')
    diffs = (
        f'=IF(E{row_number}<>"Both","",SUMPRODUCT(LEN($H{row_number}:$H{row_number})-'
        f'LEN(SUBSTITUTE($H{row_number}:$H{row_number},"D",""))))')
    left = (
        f'IF(ISBLANK(INDEX(TSMIS!C:C,$C{row_number})),"",'
        f'TRIM(INDEX(TSMIS!C:C,$C{row_number})))')
    right = (
        f'IF(ISBLANK(INDEX(TSN!C:C,$D{row_number})),"",'
        f'TRIM(INDEX(TSN!C:C,$D{row_number})))')
    value = (
        f'=IF($E{row_number}="TSMIS only",{left},'
        f'IF($E{row_number}="TSN only",{right},'
        f'IF(MID($H{row_number},1,1)="D",'
        f'IF({left}="","(blank)",{left})&" ≠ "&'
        f'IF({right}="","(blank)",{right}),{left})))')
    state = (
        f'=IF($E{row_number}<>"Both",REPT("U",1),'
        f'IF(EXACT({left},{right}),"E","D"))')
    return (*link, status, diffs, value, state)


def _read_product_input_sheet(
        ws, expected_order: Sequence[str], expected_counts: dict[str, int],
        ordinal_by_category: dict[str, int], label: str) -> dict[str, int]:
    expected_header = (
        "Comparison row", "Category", "Count", "Key (helper)",
        "__CMP_E2_BUILD_FRESH_V1_C001_B_D")
    if tuple(ws.cell(1, column).value for column in range(1, 6)) != expected_header:
        raise AuditError(f"product {label} input header drift")
    observed = {}
    for input_row, category in enumerate(expected_order, 2):
        if ws.cell(input_row, 2).value != category:
            raise AuditError(f"product {label} row {input_row} category drift")
        count = _product_int(
            ws.cell(input_row, 3).value, f"product {label} row {input_row}")
        if count != expected_counts[category]:
            raise AuditError(
                f"product {label} {category!r}: {count} != {expected_counts[category]}")
        key = f"__CMP_E2_KEY_V1_{ordinal_by_category[category]:08d}"
        if ws.cell(input_row, 4).value != key:
            raise AuditError(f"product {label} helper key drift at row {input_row}")
        observed[category] = count
    return observed


def _semantic_gaps(product_rows: Sequence[dict[str, object]],
                   truth_rows: Sequence[dict[str, object]]) -> list[dict[str, object]]:
    if len(product_rows) != len(truth_rows):
        return [{"id": "comparison_row_count_mismatch",
                 "expected": len(truth_rows), "observed": len(product_rows)}]
    gaps = []
    for expected, observed in zip(truth_rows, product_rows, strict=True):
        for field in ("category", "status", "tsmis", "tsn"):
            if observed.get(field) != expected.get(field):
                gaps.append({
                    "id": "comparison_row_mismatch", "field": field,
                    "category": expected.get("category"),
                    "expected": expected.get(field), "observed": observed.get(field),
                })
    return gaps


EXPECTED_GAP_PROBES = (
    "missing_total_accepted_as_zero",
    "all_zero_categories_with_nonzero_total_accepted",
    "boolean_and_fraction_counts_coerced",
    "duplicate_tsn_category_summed",
    "dropped_route_accepted_without_universe_diagnostic",
    "duplicate_route_accepted_and_double_counted",
    "orphan_outside_city_defaults_to_rural",
    "countless_urban_parent_is_ignored",
    "distinct_j_p_fold_is_permitted",
    "repeated_j_is_silently_summed",
)


def _inspect_production_comparison(
        formulas_path: Path, values_path: Path, truth: dict[str, object],
        tsmis_aggregate: dict[str, int], tsn_counts: dict[str, int],
        helper: dict[str, object]) -> dict[str, object]:
    expected_sheets = [
        "Summary", "Spot Check", "Comparison", "Only in TSMIS", "Only in TSN",
        "TSMIS", "TSN", "Summary by Category",
        "__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B",
    ]
    formulas = load_workbook(formulas_path, read_only=False, data_only=False)
    values = load_workbook(values_path, read_only=False, data_only=False)
    try:
        if formulas.sheetnames != expected_sheets or values.sheetnames != expected_sheets:
            raise AuditError(
                "product comparison sheet universe drift: "
                f"formulas={formulas.sheetnames!r} values={values.sheetnames!r}")
        ordinal = {row["category"]: index
                   for index, row in enumerate(truth["rows"], 1)}
        product_tsmis = _read_product_input_sheet(
            values["TSMIS"], TSMIS_ORDER, tsmis_aggregate, ordinal, "TSMIS")
        product_tsn = _read_product_input_sheet(
            values["TSN"], TSN_ORDER, tsn_counts, ordinal, "TSN")
        # Formula flavor must carry the same exact source/helper projection.
        _read_product_input_sheet(
            formulas["TSMIS"], TSMIS_ORDER, tsmis_aggregate, ordinal, "formulas TSMIS")
        _read_product_input_sheet(
            formulas["TSN"], TSN_ORDER, tsn_counts, ordinal, "formulas TSN")

        for workbook, flavor in ((formulas, "formulas"), (values, "values")):
            for input_name, snapshot_name, order in (
                    ("TSMIS", "__CMP_E2_SNAPSHOT_A", TSMIS_ORDER),
                    ("TSN", "__CMP_E2_SNAPSHOT_B", TSN_ORDER)):
                snapshot = workbook[snapshot_name]
                if snapshot.sheet_state != "veryHidden":
                    raise AuditError(f"product {flavor} {snapshot_name} visibility drift")
                source = workbook[input_name]
                for row_number in range(1, len(order) + 2):
                    if tuple(source.cell(row_number, column).value for column in range(2, 5)) \
                            != tuple(snapshot.cell(row_number, column).value
                                     for column in range(2, 5)):
                        raise AuditError(
                            f"product {flavor} {snapshot_name} row {row_number} drift")

        comparison_values = values["Comparison"]
        comparison_formulas = formulas["Comparison"]
        expected_header = (
            "Category", "#", "TSMIS Row", "TSN Row", "Status", "Diffs", "Count",
            "__CMP_E1_STATE_V1_C001_P0000_P0000")
        for ws, flavor in ((comparison_values, "values"),
                           (comparison_formulas, "formulas")):
            if (ws.max_row, ws.max_column) != (67, 8):
                raise AuditError(
                    f"product {flavor} Comparison dimensions {ws.max_row}x{ws.max_column}")
            if tuple(ws.cell(1, column).value for column in range(1, 9)) != expected_header:
                raise AuditError(f"product {flavor} Comparison header drift")

        tsmis_input_row = {category: row for row, category in enumerate(TSMIS_ORDER, 2)}
        tsn_input_row = {category: row for row, category in enumerate(TSN_ORDER, 2)}
        product_rows = []
        formula_mismatches = []
        for row_number, expected in enumerate(truth["rows"], 2):
            category = expected["category"]
            if comparison_values.cell(row_number, 1).value != category \
                    or comparison_formulas.cell(row_number, 1).value != category:
                raise AuditError(f"product Comparison category row {row_number} drift")
            expected_formulas = _expected_comparison_formulas(row_number, row_number - 1)
            observed_formulas = tuple(
                comparison_formulas.cell(row_number, column).value
                for column in range(3, 9))
            if observed_formulas != expected_formulas:
                formula_mismatches.append({
                    "row": row_number, "expected": expected_formulas,
                    "observed": observed_formulas})
            left = expected["tsmis"]
            right = expected["tsn"]
            status = expected["status"]
            diffs = int(left != right) if status == "Both" else None
            state = "D" if diffs else "E" if status == "Both" else "U"
            display = (
                f"{left} ≠ {right}" if diffs
                else str(left if left is not None else right))
            expected_values = (
                f'=HYPERLINK("#TSMIS!{tsmis_input_row[category]}:'
                f'{tsmis_input_row[category]}",{tsmis_input_row[category]})',
                (f'=HYPERLINK("#TSN!{tsn_input_row[category]}:'
                 f'{tsn_input_row[category]}",{tsn_input_row[category]})')
                if category in tsn_input_row else None,
                status, diffs, display, state,
            )
            observed_values = tuple(
                comparison_values.cell(row_number, column).value
                for column in range(3, 9))
            if observed_values != expected_values:
                raise AuditError(
                    f"product values Comparison row {row_number} drift: "
                    f"{observed_values!r} != {expected_values!r}")
            if any(comparison_values.cell(row_number, column).data_type == "f"
                   for column in range(5, 9)):
                raise AuditError(
                    f"product values flavor retained verdict formula row {row_number}")
            product_rows.append({
                "category": category, "status": status,
                "tsmis": product_tsmis.get(category),
                "tsn": product_tsn.get(category),
                "diffs": diffs, "state": state,
            })
        if formula_mismatches:
            raise AuditError(
                f"product formula contract drift: {formula_mismatches[:1]!r}")

        only_tsmis = values["Only in TSMIS"]
        only_tsn = values["Only in TSN"]
        if (only_tsmis.max_row, only_tsmis.max_column) != (9, 4):
            raise AuditError("product Only in TSMIS shape drift")
        if tuple(only_tsmis.cell(row, 1).value for row in range(2, 10)) \
                != EXPECTED_ONLY_TSMIS:
            raise AuditError("product Only in TSMIS category order drift")
        if (only_tsn.max_row, only_tsn.max_column) != (1, 4):
            raise AuditError("product Only in TSN shape drift")

        familiar = values["Summary by Category"]
        if (familiar.max_row, familiar.max_column) != (83, 4):
            raise AuditError(
                f"product familiar dimensions {familiar.max_row}x{familiar.max_column}")
        familiar_rows = []
        row_number = 7
        for section, _source, cats in SECTIONS:
            if familiar.cell(row_number, 1).value != section:
                raise AuditError(f"product familiar section row {row_number} drift")
            row_number += 1
            for cat in cats:
                left = tsmis_aggregate[cat.category]
                right = tsn_counts.get(cat.category)
                observed = tuple(
                    familiar.cell(row_number, column).value for column in range(1, 5))
                expected = (
                    cat.display_label, left, right,
                    right - left if right is not None else None)
                if observed != expected:
                    raise AuditError(
                        f"product familiar row {row_number}: {observed!r} != {expected!r}")
                familiar_rows.append((section, *expected))
                row_number += 1
        if row_number != 82 or any(
                familiar.cell(82, column).value is not None for column in range(1, 5)):
            raise AuditError("product familiar pre-total row drift")
        if tuple(familiar.cell(83, column).value for column in range(1, 5)) != (
                "Total Intersections", EXPECTED_TSMIS_TOTAL, EXPECTED_TSN_TOTAL, 167):
            raise AuditError("product familiar total drift")
        note = str(familiar["A4"].value or "")
        classification_note = str(familiar["A2"].value or "")
        note_claims = {
            "jp_fold_disclosed": "J–P" in note and "folded" in note,
            "tsnr_reference_disclosed": "TSNR/MIRE reference" in note,
            "roundabout_one_sided_disclosed": "Roundabout (R) stays one-sided" in note,
        }
        if not all(note_claims.values()):
            raise AuditError(f"product familiar decision note drift: {note_claims!r}")
        misleading_classification_note = {
            "claims_unclassified_categories_show_zero": (
                "show 0 on that side" in classification_note),
            "uses_irrelevant_ramp_pv_example": (
                "ramp types P / V" in classification_note),
            "actual_tsmis_only_cells_are_blank": all(
                familiar.cell(row, 3).value is None
                and familiar.cell(row, 4).value is None
                for row in (25, 26, 27, 29, 44, 46, 47, 69)),
        }
        if not all(misleading_classification_note.values()):
            raise AuditError(
                "product familiar known classification-note defect drift: "
                f"{misleading_classification_note!r}")

        summary = values["Summary"]
        summary_counts = {
            "tsmis_rows": summary["C8"].value,
            "tsn_rows": summary["C9"].value,
            "union_rows": summary["C10"].value,
            "paired_rows": summary["C12"].value,
            "tsmis_only_rows": summary["C13"].value,
            "tsn_only_rows": summary["C14"].value,
            "differing_rows": summary["C16"].value,
            "identical_rows": summary["C17"].value,
            "differing_cells": summary["C18"].value,
            "count_field_differences": summary["D22"].value,
        }
        expected_summary_counts = {
            "tsmis_rows": 66, "tsn_rows": 58, "union_rows": 66,
            "paired_rows": 58, "tsmis_only_rows": 8, "tsn_only_rows": 0,
            "differing_rows": 53, "identical_rows": 5,
            "differing_cells": 53, "count_field_differences": 53,
        }
        if summary_counts != expected_summary_counts:
            raise AuditError(f"product Summary literal count drift: {summary_counts!r}")

        strings = _workbook_strings(values)
        provenance_tokens = (
            "07/09/2026", "07/10/2026", "Yunus.Shaikh@dot.ca.gov",
            "OTM22250", "09/15/2025", "TRLBUGNI", "4843738", "04:53 PM",
            "F-FOUR WAY FLASHER (RED ON ALL)",
        )
        provenance_presence = {
            token: any(token in value for value in strings)
            for token in provenance_tokens
        }
        source_note = str(summary["B5"].value or "")
        familiar_source_note = str(familiar["A3"].value or "")
    finally:
        formulas.close()
        values.close()

    gaps = _semantic_gaps(product_rows, truth["rows"])
    if gaps:
        raise AuditError(f"product comparison semantic gaps: {gaps[:3]!r}")
    comparison_result = helper.get("comparison")
    consolidation_result = helper.get("consolidation")
    if not isinstance(comparison_result, dict) or not isinstance(
            consolidation_result, dict):
        raise AuditError("product helper outcome shape drift")
    expected_counts = {
        "known": True, "paired_rows": 58, "side_a_only_rows": 8,
        "side_b_only_rows": 0, "differing_rows": 53,
        "differing_cells": 53, "asserted_cells": 58,
        "context_cells": 0, "per_field_counts": {"1:Count": 53},
    }
    observed_counts = comparison_result.get("counts")
    if not isinstance(observed_counts, dict) or any(
            observed_counts.get(key) != value for key, value in expected_counts.items()):
        raise AuditError(f"product structured comparison counts drift: {observed_counts!r}")
    if (comparison_result.get("status"), comparison_result.get("completion"),
            comparison_result.get("verdict"), comparison_result.get("skipped_inputs"),
            comparison_result.get("failed_inputs")) != ("ok", "complete", "diff", 0, 0):
        raise AuditError("product structured comparison outcome drift")
    if comparison_result.get("warnings") or comparison_result.get("failures"):
        raise AuditError("product comparison unexpectedly reported warnings/failures")
    if (consolidation_result.get("status"), consolidation_result.get("completion"),
            consolidation_result.get("skipped_inputs"),
            consolidation_result.get("failed_inputs")) != ("ok", "complete", 0, 0):
        raise AuditError("product consolidation outcome drift")
    probes = helper.get("known_product_gap_probes")
    if not isinstance(probes, dict) or probes.get(
            "all_expected_product_red_paths_reproduced") is not True:
        raise AuditError("product known-gap probe summary missing/not all reproduced")
    probe_values = probes.get("probes")
    if (not isinstance(probe_values, dict)
            or tuple(probe_values) != EXPECTED_GAP_PROBES
            or not all(value is True for value in probe_values.values())
            or probes.get("baseline_total") != 16_459
            or probes.get("dropped_route_total") != 16_458
            or probes.get("duplicate_route_total") != 17_752):
        raise AuditError(f"product known-gap probe contract drift: {probes!r}")

    formula_semantics = _workbook_semantic_digest(formulas_path)
    value_semantics = _workbook_semantic_digest(values_path)
    if formula_semantics["formula_cells"] <= value_semantics["formula_cells"]:
        raise AuditError("product formulas/values flavor formula census inverted")
    return {
        "formulas": {
            "package_without_core": _zip_digest_without_core(formulas_path),
            "semantic_digest": formula_semantics,
        },
        "values": {
            "package_without_core": _zip_digest_without_core(values_path),
            "semantic_digest": value_semantics,
        },
        "product_rows": product_rows,
        "structured_counts": expected_counts,
        "semantic_gaps": [], "comparison_semantics_exact": True,
        "formula_rows_exact": len(product_rows),
        "familiar_sheet": {
            "numeric_rows_exact": True, "category_rows": len(CATEGORIES),
            "ordered_typed_sha256": _canonical_digest(familiar_rows),
            "decision_note_claims": note_claims,
            "classification_note": classification_note,
            "misleading_classification_note": misleading_classification_note,
            "classification_note_contradicts_one_sided_blanks": True,
        },
        "known_product_gap_probes": probes,
        "provenance": {
            "summary_source_note": source_note,
            "familiar_source_note": familiar_source_note,
            "saved_notes_use_basenames_only": (
                "intersection_summary_consolidated.xlsx" in source_note
                and "tsn_intersection_summary_normalized.xlsx" in source_note
                and "\\" not in source_note and "/" not in source_note),
            "printed_source_provenance_presence": provenance_presence,
            "printed_source_provenance_preserved": any(provenance_presence.values()),
        },
        "source_backed_values_exact": True,
    }


def _sanitized_helper_result(payload: dict[str, object]) -> dict[str, object]:
    comparison = payload["comparison"]
    consolidation = payload["consolidation"]
    generation = comparison.get("artifact_generation") or {}
    members = sorted(({
        "flavor": member.get("flavor"),
        "commit_role": member.get("commit_role"),
        "filename": Path(str(member.get("path"))).name,
    } for member in generation.get("members", [])), key=lambda item: str(item["flavor"]))
    return {
        "consolidation": {key: consolidation.get(key) for key in (
            "status", "completion", "skipped_inputs", "failed_inputs")},
        "comparison": {key: comparison.get(key) for key in (
            "status", "completion", "verdict", "skipped_inputs", "failed_inputs",
            "counts", "warnings", "failures")},
        "artifact_generation": {
            "completion": generation.get("completion"),
            "publication_state": generation.get("publication_state"),
            "requested_mode": generation.get("requested_mode"), "members": members,
        },
        "known_product_gap_probes": payload.get("known_product_gap_probes"),
    }


def _parse_helper_stdout(stdout: str) -> dict[str, object]:
    for line in reversed([line.strip() for line in stdout.splitlines() if line.strip()]):
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict) and "outputs" in payload:
            return payload
    raise AuditError("product helper emitted no structured JSON result")


def _run_product_once(
        run_root: Path, xlsx_root: Path, tsn_xlsx: Path,
        raw_records: dict[str, dict[str, int]], truth: dict[str, object],
        tsmis_aggregate: dict[str, int], tsn_counts: dict[str, int]
        ) -> dict[str, object]:
    completed = subprocess.run(
        [
            sys.executable, str(PRODUCT_HELPER_PATH),
            "--xlsx-root", str(xlsx_root), "--tsn-xlsx", str(tsn_xlsx),
            "--work-root", str(run_root),
        ],
        cwd=REPO_ROOT, env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        text=True, encoding="utf-8", errors="replace", capture_output=True,
        timeout=300, check=False)
    if completed.returncode != 0:
        raise AuditError(
            f"product helper failed ({completed.returncode}): "
            f"{completed.stderr[-2000:]!r} {completed.stdout[-2000:]!r}")
    payload = _parse_helper_stdout(completed.stdout)
    outputs = payload.get("outputs")
    if not isinstance(outputs, dict):
        raise AuditError("product helper output path map missing")
    expected_paths = {
        "consolidated": run_root / "intersection_summary_consolidated.xlsx",
        "formulas": run_root / "intersection_summary_comparison.xlsx",
        "values": run_root / "intersection_summary_comparison (values).xlsx",
    }
    for label, expected in expected_paths.items():
        observed = Path(str(outputs.get(label))).resolve()
        if observed != expected.resolve() or not observed.is_file():
            raise AuditError(f"product helper {label} path escaped/drifted: {observed}")
    consolidated = _inspect_production_consolidated(
        expected_paths["consolidated"], raw_records, tsmis_aggregate)
    comparison = _inspect_production_comparison(
        expected_paths["formulas"], expected_paths["values"], truth,
        tsmis_aggregate, tsn_counts, payload)
    product_code = payload.get("loaded_product_code")
    if not isinstance(product_code, dict) or not product_code.get("entries"):
        raise AuditError("product helper loaded-code manifest missing")
    return {
        "helper_result": _sanitized_helper_result(payload),
        "consolidated": consolidated, "comparison": comparison,
        "loaded_product_code": product_code,
    }


def _run_product(
        work_root: Path, xlsx_root: Path, tsn_xlsx: Path,
        raw_records: dict[str, dict[str, int]], truth: dict[str, object],
        tsmis_aggregate: dict[str, int], tsn_counts: dict[str, int]
        ) -> dict[str, object]:
    first = _run_product_once(
        work_root / "product_a", xlsx_root, tsn_xlsx, raw_records, truth,
        tsmis_aggregate, tsn_counts)
    second = _run_product_once(
        work_root / "product_b", xlsx_root, tsn_xlsx, raw_records, truth,
        tsmis_aggregate, tsn_counts)
    if first != second:
        left = _sha_bytes(json.dumps(
            first, sort_keys=True, ensure_ascii=False,
            separators=(",", ":")).encode("utf-8"))
        right = _sha_bytes(json.dumps(
            second, sort_keys=True, ensure_ascii=False,
            separators=(",", ":")).encode("utf-8"))
        raise AuditError(f"production stable replay differs: {left} != {right}")
    return {
        "replays": 2, "stable_replay_exact": True,
        "stable_replay_sha256": _sha_bytes(json.dumps(
            first, sort_keys=True, ensure_ascii=False,
            separators=(",", ":")).encode("utf-8")),
        **first,
    }


def _resolve_rural_urban_rows(
        rows: Sequence[tuple[object, str]]) -> dict[str, int]:
    """Strict independent parent resolver used by the permanent mutation gate."""
    result: dict[str, int] = {}
    parent: str | None = None
    for value, raw_label in rows:
        label = str(raw_label).strip().upper()
        if label.startswith("R-RURAL"):
            parent = "R"
            if value is not None:
                result["R"] = _strict_count(value, "rural parent")
        elif label.startswith("U-URBAN"):
            parent = "U"
            if value is not None:
                result["U"] = _strict_count(value, "urban parent")
        elif label.startswith("-O"):
            if parent is None:
                raise AuditError("orphan -O OUTSIDE CITY row")
            key = f"{parent}-O"
            if key in result:
                raise AuditError(f"duplicate rural/urban child {key}")
            result[key] = _strict_count(value, f"{key} child")
        elif label.startswith("+"):
            if "+" in result:
                raise AuditError("duplicate rural/urban + row")
            result["+"] = _strict_count(value, "rural/urban +")
        else:
            raise AuditError(f"unknown rural/urban row {raw_label!r}")
    return result


def _fold_tsn_control_rows(
        rows: Sequence[tuple[str, object]]) -> dict[str, int]:
    """Permit distinct J/K/L/M/N/P source rows but reject exact duplicates."""
    result: dict[str, int] = {}
    seen = set()
    for raw_code, raw_count in rows:
        code = str(raw_code).upper()
        if code in seen:
            raise AuditError(f"duplicate TSN control source code {code}")
        seen.add(code)
        count = _strict_count(raw_count, f"TSN control {code}")
        target = "S" if code in set("JKLMNP") else code
        result[target] = result.get(target, 0) + count
    return result


def _loaded_oracle_module_manifest() -> dict[str, object]:
    entries = []
    for module_name, module in sorted(sys.modules.items()):
        if module_name.split(".", 1)[0] not in {
                "pdfplumber", "pdfminer", "openpyxl"}:
            continue
        raw = getattr(module, "__file__", None)
        if not raw:
            continue
        path = Path(raw).resolve()
        if path.suffix.lower() != ".py" or not path.is_file():
            continue
        entries.append({
            "module": module_name, "filename": path.name,
            "bytes": path.stat().st_size, "sha256": _sha_file(path),
        })
    canonical = json.dumps(
        entries, sort_keys=True, ensure_ascii=False,
        separators=(",", ":")).encode("utf-8")
    return {
        "module_file_count": len(entries),
        "canonical_json_sha256": _sha_bytes(canonical), "entries": entries,
    }


def _product_manifest_current(manifest: dict[str, object]) -> dict[str, object]:
    checks = []
    entries = manifest.get("entries")
    if not isinstance(entries, list):
        return {"all_current": False, "checks": [], "reason": "entries missing"}
    for entry in entries:
        relative = str(entry.get("relative_path", ""))
        path = (REPO_ROOT / "scripts" / relative).resolve()
        expected = {"bytes": entry.get("bytes"), "sha256": entry.get("sha256")}
        observed = _file_identity(path) if path.is_file() else None
        checks.append({
            "module": entry.get("module"), "relative_path": relative,
            "expected": expected, "observed": observed,
            "current": observed == expected,
        })
    return {
        "all_current": bool(checks) and all(check["current"] for check in checks),
        "checks": checks,
    }


def _run_gate(path: Path) -> dict[str, object]:
    completed = subprocess.run(
        [sys.executable, str(path)], cwd=REPO_ROOT,
        text=True, encoding="utf-8", errors="replace", capture_output=True,
        timeout=240, check=False,
        env={**os.environ, "PYTHONIOENCODING": "utf-8"})
    if completed.returncode != 0:
        raise AuditError(
            f"mutation gate failed ({completed.returncode}): "
            f"{completed.stdout[-3000:]!r} {completed.stderr[-3000:]!r}")
    return {
        "status": "executed_pass", "gate_identity": _file_identity(path),
        "stdout": completed.stdout.strip(), "stderr": completed.stderr.strip(),
    }


def _mutation_probes(truth: dict[str, object]) -> dict[str, object]:
    probes = []

    def record(name: str, detected: bool) -> None:
        probes.append({"name": name, "detected": bool(detected)})

    for label, value in (("boolean", True), ("float", 1.0),
                         ("string", "1"), ("negative", -1)):
        try:
            _strict_count(value, f"probe {label}")
        except AuditError:
            record(f"strict_count_rejects_{label}", True)
        else:
            record(f"strict_count_rejects_{label}", False)
    record("suffix_route_identity_preserved",
           _route_from_name(Path("intersection_summary_route_008U.pdf"), ".pdf")
           == "008U")
    changed = _cross_format(
        {"001": {key: value for key, value in EXPECTED_TSMIS_AGGREGATE.items()}},
        {"001": {**EXPECTED_TSMIS_AGGREGATE,
                 TSMIS_ORDER[0]: EXPECTED_TSMIS_AGGREGATE[TSMIS_ORDER[0]] + 1}})
    record("cross_format_single_cell_mutation", changed["difference_count"] == 1)
    record("valid_countless_urban_parent_retained",
           _resolve_rural_urban_rows([
               (None, "U-URBAN -I INSIDE CITY"),
               (5, "-O OUTSIDE CITY")]).get("U-O") == 5)
    try:
        _resolve_rural_urban_rows([(5, "-O OUTSIDE CITY")])
    except AuditError:
        record("orphan_outside_city_rejected", True)
    else:
        record("orphan_outside_city_rejected", False)
    record("distinct_j_p_fold_permitted",
           _fold_tsn_control_rows([("J", 1), ("P", 2)]) == {"S": 3})
    try:
        _fold_tsn_control_rows([("J", 1), ("J", 2)])
    except AuditError:
        record("repeated_j_rejected", True)
    else:
        record("repeated_j_rejected", False)
    corrected = [{
        "category": row["category"], "status": row["status"],
        "tsmis": row["tsmis"], "tsn": row["tsn"],
    } for row in truth["rows"]]
    record("exact_comparison_truth_has_no_semantic_gaps",
           not _semantic_gaps(corrected, truth["rows"]))
    wrong = [dict(row) for row in corrected]
    wrong[0]["tsmis"] += 1
    record("comparison_value_mutation_detected",
           bool(_semantic_gaps(wrong, truth["rows"])))
    return {"probes": probes,
            "all_detected": all(probe["detected"] for probe in probes)}


def _publication_current(args: argparse.Namespace,
                         result: dict[str, object]) -> tuple[bool, dict[str, object]]:
    tree_paths = {
        "tsmis_xlsx": args.summary_xlsx_root,
        "tsmis_pdf": args.summary_pdf_root,
    }
    file_paths = {
        "tsn_raw_pdf": args.tsn_raw,
        "tsn_normalized_xlsx": args.tsn_xlsx,
        "stage6_result": args.stage6_result,
        "stage6_acceptance": args.stage6_acceptance,
        "tsn_cross_format_result": args.cross_format_result,
        "tsnr_reference": args.tsnr_reference,
    }
    detail: dict[str, object] = {"trees": {}, "files": {}, "code": {}}
    captures = result["source_capture"]
    for label, path in tree_paths.items():
        observed, _entries = _manifest(path, str(TREE_BINDINGS[label]["suffix"]))
        expected = captures[label]["observed"]
        detail["trees"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    for label, path in file_paths.items():
        observed = _file_identity(path) if path.is_file() else None
        expected = captures[label]["observed"]
        detail["files"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    code_paths = {
        "generator": GENERATOR_PATH, "product_helper": PRODUCT_HELPER_PATH,
        "self_gate": SELF_GATE_PATH,
    }
    for label, path in code_paths.items():
        observed = _file_identity(path) if path.is_file() else None
        expected = result["provenance"]["code_identities"][label]
        detail["code"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    parser_manifest = _loaded_oracle_module_manifest()
    expected_parser = result["provenance"]["loaded_oracle_module_manifest"]
    detail["code"]["oracle_parser_modules"] = {
        "expected_sha256": expected_parser["canonical_json_sha256"],
        "observed_sha256": parser_manifest["canonical_json_sha256"],
        "expected_files": expected_parser["module_file_count"],
        "observed_files": parser_manifest["module_file_count"],
        "current": parser_manifest == expected_parser,
    }
    product_current = _product_manifest_current(
        result["production"]["loaded_product_code"])
    detail["code"]["loaded_product_modules"] = product_current
    flags = [
        item["current"]
        for group in (detail["trees"], detail["files"], detail["code"])
        for item in group.values() if "current" in item
    ]
    flags.append(product_current["all_current"])
    return all(flags), detail


def _assert_frozen_digest(label: str, observed: str, expected: str) -> None:
    if expected == "TO_BE_FROZEN":
        raise AuditError(f"unfrozen {label} observed digest: {observed}")
    if observed != expected:
        raise AuditError(f"{label} digest drift: {observed} != {expected}")


def run(args: argparse.Namespace) -> dict[str, object]:
    args.work_root.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
            prefix="source-transaction-", dir=args.work_root) as raw_transaction:
        transaction = Path(raw_transaction)
        source_capture: dict[str, object] = {}
        source_capture["tsmis_xlsx"], xlsx_root = _capture_tree(
            "tsmis_xlsx", args.summary_xlsx_root, TREE_BINDINGS["tsmis_xlsx"],
            transaction / "tsmis_xlsx")
        source_capture["tsmis_pdf"], pdf_root = _capture_tree(
            "tsmis_pdf", args.summary_pdf_root, TREE_BINDINGS["tsmis_pdf"],
            transaction / "tsmis_pdf")
        file_specs = (
            ("tsn_raw_pdf", args.tsn_raw, "tsn_raw.pdf"),
            ("tsn_normalized_xlsx", args.tsn_xlsx, "tsn_normalized.xlsx"),
            ("stage6_result", args.stage6_result, "stage6_result.json"),
            ("stage6_acceptance", args.stage6_acceptance, "stage6_acceptance.json"),
            ("tsn_cross_format_result", args.cross_format_result, "cross_format.json"),
            ("tsnr_reference", args.tsnr_reference, "tsnr_reference.xlsx"),
        )
        private_files = {}
        for label, source, filename in file_specs:
            source_capture[label], private_files[label] = _capture_file(
                label, source, FILE_BINDINGS[label], transaction / filename)

        tsmis_xlsx = _parse_tsmis_xlsx(xlsx_root)
        tsmis_pdf = _parse_tsmis_pdfs(pdf_root)
        routes = _require_exact_route_universes({
            "tsmis_xlsx": tsmis_xlsx["routes"],
            "tsmis_pdf": tsmis_pdf["routes"],
        })
        cross_format = _cross_format(
            tsmis_xlsx["records"], tsmis_pdf["records"])
        aggregate = _aggregate(tsmis_xlsx["records"])
        if aggregate != EXPECTED_TSMIS_AGGREGATE:
            changed = [key for key in TSMIS_ORDER
                       if aggregate.get(key) != EXPECTED_TSMIS_AGGREGATE.get(key)]
            raise AuditError(f"TSMIS exact aggregate drift: {changed[:5]!r}")
        aggregate_digest = _canonical_digest(
            (key, aggregate[key]) for key in TSMIS_ORDER)
        tsn = _load_tsn_normalized(private_files["tsn_normalized_xlsx"])
        tsn_raw = _load_tsn_raw_pdf(private_files["tsn_raw_pdf"])
        stage6 = _load_stage6(
            private_files["stage6_result"], private_files["stage6_acceptance"])
        tsn_cross_format = _load_cross_format_result(
            private_files["tsn_cross_format_result"])
        tsnr_reference = _load_tsnr_reference(private_files["tsnr_reference"])
        truth = _comparison_truth(aggregate, tsn["counts"])

        frozen = {
            "tsmis_cross_format": (
                cross_format["ordered_typed_sha256"],
                EXPECTED_TSMIS_CROSS_FORMAT_SHA256),
            "tsmis_aggregate": (
                aggregate_digest, EXPECTED_TSMIS_AGGREGATE_SHA256),
            "comparison_truth": (
                truth["ordered_typed_sha256"], EXPECTED_COMPARISON_TRUTH_SHA256),
        }
        unfrozen = {label: observed for label, (observed, expected) in frozen.items()
                    if expected == "TO_BE_FROZEN"}
        if unfrozen:
            raise AuditError(
                "unfrozen Stage-8 digests: " + json.dumps(unfrozen, sort_keys=True))
        for label, (observed, expected) in frozen.items():
            _assert_frozen_digest(label, observed, expected)

        product = _run_product(
            transaction / "product", xlsx_root,
            private_files["tsn_normalized_xlsx"], tsmis_xlsx["records"], truth,
            aggregate, tsn["counts"])

    mutations = _mutation_probes(truth)
    mutation_gate = _run_gate(SELF_GATE_PATH)
    code_identities = {
        "generator": _file_identity(GENERATOR_PATH),
        "product_helper": _file_identity(PRODUCT_HELPER_PATH),
        "self_gate": _file_identity(SELF_GATE_PATH),
    }
    parser_manifest = _loaded_oracle_module_manifest()
    product_code_current = _product_manifest_current(product["loaded_product_code"])

    source_invariants = {
        "exact_bound_tsmis_tree_captures": all(
            source_capture[label]["observed"][key] == TREE_BINDINGS[label][key]
            for label in ("tsmis_xlsx", "tsmis_pdf")
            for key in ("files", "bytes", "manifest_sha256")),
        "exact_217_route_universe_order_suffixes_and_170_absence": routes["all_exact"],
        "all_217_xlsx_fixed_layouts_and_non_highway_partitions_exact": (
            tsmis_xlsx["files"] == 217
            and tsmis_xlsx["source_value_cells"] == 14_322
            and tsmis_xlsx["all_non_highway_partitions_exact"]),
        "all_217_pdf_fixed_geometries_provenance_and_partitions_exact": (
            tsmis_pdf["files"] == 217 and tsmis_pdf["pages"] == 434
            and tsmis_pdf["source_value_cells"] == 14_322
            and tsmis_pdf["metadata_signature_count"] == 173),
        "all_14322_tsmis_excel_pdf_values_exact": (
            cross_format["all_exact"] and cross_format["paired_values"] == 14_322),
        "exact_tsmis_aggregate_and_total": (
            aggregate == EXPECTED_TSMIS_AGGREGATE
            and aggregate["Total Intersections"] == EXPECTED_TSMIS_TOTAL),
        "exact_tsn_raw_pdf_identity_provenance_and_printed_f_g": (
            tsn_raw["report_provenance"]["total_intersections"] == EXPECTED_TSN_TOTAL
            and tsn_raw["raw_control_f_label"]
            == "F-FOUR WAY FLASHER (RED ON ALL)"),
        "exact_tsn_r7_normalized_order_types_and_values": (
            tsn["rows"] == 58 and tuple(tsn["order"]) == TSN_ORDER
            and tsn["counts"]["Total Intersections"] == EXPECTED_TSN_TOTAL),
        "accepted_stage6_chain_exact_projection_and_false_full_conservation": (
            stage6["projection_exact"]
            and stage6["stage6_family_audit_complete"]
            and stage6["normalized_full_conservation"] is False),
        "accepted_tsn_detail_xlsx_pdf_summary_cross_format_chain_exact": (
            tsn_cross_format["status"] == "pass"
            and tsn_cross_format["unresolved_gate_count"] == 0
            and tsn_cross_format["all_report_view_columns_mapped_once"]),
        "tsnr_f_g_and_jp_mapping_semantics_proven": (
            tsnr_reference["semantic_mapping_proven"]
            and tsnr_reference["jp_members"] == list("JKLMNP")),
        "exact_66_row_comparison_truth": (
            truth["counts"]["union_rows"] == 66
            and truth["counts"]["paired_rows"] == 58
            and truth["counts"]["tsmis_only_rows"] == 8
            and truth["counts"]["differing_shared_rows"] == 53),
    }
    audit_invariants = {
        **source_invariants,
        "production_two_replays_semantically_and_package_stable": (
            product["replays"] == 2 and product["stable_replay_exact"]),
        "production_source_backed_consolidated_and_comparison_values_exact": (
            product["consolidated"]["projection_exact"]
            and product["comparison"]["source_backed_values_exact"]),
        "production_formula_and_values_twin_semantics_exact": (
            product["comparison"]["formula_rows_exact"] == 66
            and product["comparison"]["formulas"]["semantic_digest"]["formula_cells"]
            > product["comparison"]["values"]["semantic_digest"]["formula_cells"]),
        "production_current_source_union_status_and_verdict_semantics_exact": (
            product["comparison"]["comparison_semantics_exact"]
            and not product["comparison"]["semantic_gaps"]),
        "all_known_product_red_paths_reproduced_in_isolated_child": (
            product["comparison"]["known_product_gap_probes"]
            ["all_expected_product_red_paths_reproduced"]),
        "permanent_semantic_and_publication_gate_passed": (
            mutation_gate["status"] == "executed_pass"),
        "local_mutation_probes_all_detected": mutations["all_detected"],
        "loaded_oracle_parser_manifest_nonempty": parser_manifest["module_file_count"] > 0,
        "loaded_product_code_current_at_result_build": product_code_current["all_current"],
        "visual_pdf_review_and_all_member_pixel_census_complete": (
            VISUAL_REVIEW["all_217_page2_pixel_census"] is True),
    }
    projection_exact = bool(
        product["consolidated"]["projection_exact"]
        and product["comparison"]["source_backed_values_exact"])
    production_semantics_exact = bool(
        product["comparison"]["comparison_semantics_exact"])

    result = {
        "schema_version": 1,
        "audit": (
            "Stage 8 Intersection Summary authoritative TSMIS-vs-TSN base "
            "comparison oracle"),
        "methodology": {
            "authority": (
                "Exact All Reports 7.9 same-pull 217-route TSMIS Intersection "
                "Summary XLSX/PDF bytes, exact accepted TSN raw-to-r7 chain, "
                "accepted TSN Detail XLSX/PDF/Summary cross-format oracle, and TSNR "
                "control/geometry decision source."),
            "independence": (
                "Truth-side parsing imports no application parser, normalizer, schema, "
                "comparator, or writer. Production imports only in an isolated child."),
            "outcomes_separated": [
                "source_truth_exact", "production_value_projection_exact",
                "production_comparison_semantics_exact",
                "normalized_source_full_conservation",
                "stage8_base_oracle_complete", "comparison_end_to_end_perfect",
            ],
            "comparison_contract": (
                "58 shared categories, eight structurally TSMIS-only categories, "
                "zero TSN-only categories; structural absence never becomes zero. "
                "J/K/L/M/N/P are source claims feeding a declared S projection."),
            "visual_verification": VISUAL_REVIEW,
        },
        "bindings": {
            "source_trees": TREE_BINDINGS, "files": FILE_BINDINGS,
            "expected_route_lf_sha256": EXPECTED_ROUTE_LF_SHA256,
            "expected_tsmis_cross_format_sha256": EXPECTED_TSMIS_CROSS_FORMAT_SHA256,
            "expected_tsmis_aggregate_sha256": EXPECTED_TSMIS_AGGREGATE_SHA256,
            "expected_comparison_truth_sha256": EXPECTED_COMPARISON_TRUTH_SHA256,
        },
        "source_capture": source_capture,
        "route_universes": routes,
        "tsmis": {
            "xlsx": {
                "routes": tsmis_xlsx["routes"],
                "records": tsmis_xlsx["records"],
                "source_value_cells": tsmis_xlsx["source_value_cells"],
                "section_arithmetic_sha256": tsmis_xlsx[
                    "section_arithmetic_sha256"],
            },
            "pdf": {key: value for key, value in tsmis_pdf.items()
                    if key != "records"},
            "cross_format": cross_format,
            "aggregate": aggregate,
            "aggregate_ordered_typed_sha256": aggregate_digest,
        },
        "tsn": {
            "raw_pdf": tsn_raw, "normalized": tsn,
            "stage6_chain": stage6,
            "detail_xlsx_pdf_summary_cross_format": tsn_cross_format,
            "tsnr_reference": tsnr_reference,
        },
        "intended_comparison_truth": truth,
        "production": product,
        "semantic_mutation_probes": mutations,
        "dependency_gates": {
            "stage8_intersection_summary_mutations": mutation_gate,
        },
        "provenance": {
            "code_identities": code_identities,
            "loaded_oracle_module_manifest": parser_manifest,
            "loaded_product_code_current_at_result_build": product_code_current,
        },
        "findings": {
            "oracle_blocking": [],
            "product_red": [
                {"finding": "CMP-AUD-020", "fact": (
                    "Missing Total and all-zero category partitions with a nonzero "
                    "Total are accepted by the aggregate loader."),
                 "evidence": product["comparison"]["known_product_gap_probes"]},
                {"finding": "CMP-AUD-021", "fact": (
                    "Boolean and fractional aggregate counts are coerced instead of "
                    "being rejected with category/source context."),
                 "evidence": product["comparison"]["known_product_gap_probes"]},
                {"finding": "CMP-AUD-022", "fact": (
                    "Repeated exact normalized categories are silently summed even "
                    "though only distinct J/K/L/M/N/P-to-S folding is authorized."),
                 "evidence": product["comparison"]["known_product_gap_probes"]},
                {"finding": "CMP-AUD-023", "fact": (
                    "An orphan -O child defaults to Rural and a count-less Urban "
                    "parent is ignored by the product block walker."),
                 "evidence": product["comparison"]["known_product_gap_probes"]},
                {"finding": "CMP-AUD-076", "fact": (
                    "Saved product artifacts retain basenames but no exact TSMIS "
                    "report date, reference date, submitter, or generation metadata."),
                 "evidence": {
                     "consolidated": product["consolidated"]
                     ["printed_tsmis_provenance_presence"],
                     "comparison": product["comparison"]["provenance"]}},
                {"finding": "CMP-AUD-144", "fact": (
                    "The accepted S=2648 comparison projection cannot reconstruct "
                    "the six raw J/K/L/M/N/P labels and counts."),
                 "evidence": stage6["signal_fold"]},
                {"finding": "CMP-AUD-145", "fact": (
                    "TSNR and same-pull TSMIS prove canonical F=RED/MAINLINE, but "
                    "normalization/comparison do not retain the TSN PDF's erroneous "
                    "raw F=RED ON ALL claim or declared correction provenance."),
                 "evidence": {
                     "raw": tsn_raw["raw_control_f_label"],
                     "canonical": tsnr_reference["control_f_canonical"],
                     "product_presence": product["comparison"]["provenance"]
                     ["printed_source_provenance_presence"]}},
                {"finding": "CMP-AUD-146", "fact": (
                    "Normalized/product artifacts omit typed TSN report identity, "
                    "dates, submitter, event, scope, and generation time."),
                 "evidence": tsn_raw["report_provenance"]},
                {"finding": "CMP-AUD-183", "fact": (
                    "Dropping route 905 is accepted at total 16458 and duplicating "
                    "route 001 is accepted at total 17752, without a route-universe "
                    "diagnostic."),
                 "evidence": product["comparison"]["known_product_gap_probes"]},
                {"finding": "CMP-AUD-184", "fact": (
                    "The familiar Intersection view correctly renders structurally "
                    "absent TSN categories as blank, but its note says they show zero "
                    "and cites an irrelevant Ramp P/V example."),
                 "evidence": product["comparison"]["familiar_sheet"]},
            ],
        },
        "source_invariants": source_invariants,
        "audit_invariants": audit_invariants,
        "source_truth_exact": all(source_invariants.values()),
        "production_value_projection_exact": projection_exact,
        "production_comparison_semantics_exact": production_semantics_exact,
        "normalized_source_full_conservation": False,
        "stage8_base_oracle_complete": all(audit_invariants.values()),
        "comparison_end_to_end_perfect": False,
    }
    current, current_detail = _publication_current(args, result)
    result["provenance"]["final_revalidation_at_result_build"] = {
        "all_current": current, "detail": current_detail}
    result["audit_invariants"]["sources_and_code_current_at_result_build"] = current
    result["source_truth_exact"] = bool(result["source_truth_exact"] and current)
    result["stage8_base_oracle_complete"] = all(
        result["audit_invariants"].values())
    return result


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
                mode="w", encoding="utf-8", newline="\n", delete=False,
                dir=path.parent, prefix=f".{path.name}.", suffix=".tmp") as handle:
            temporary = Path(handle.name)
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary is not None:
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass


def _unlink_if_present(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass


def _write_decision(path: Path, output: Path, result: dict[str, object],
                    *, accepted: bool, reason: str,
                    postwrite_current: bool,
                    postwrite_detail: dict[str, object],
                    open_findings_authorized: bool) -> dict[str, object]:
    result_identity = _file_identity(output)
    decision = {
        "schema_version": 1, "accepted": accepted, "reason": reason,
        "audit": result.get("audit"), "result": str(output.resolve()),
        "result_bytes": result_identity["bytes"],
        "result_sha256": result_identity["sha256"],
        "source_truth_exact": result.get("source_truth_exact", False),
        "production_value_projection_exact": result.get(
            "production_value_projection_exact", False),
        "production_comparison_semantics_exact": result.get(
            "production_comparison_semantics_exact", False),
        "normalized_source_full_conservation": result.get(
            "normalized_source_full_conservation", False),
        "stage8_base_oracle_complete": result.get(
            "stage8_base_oracle_complete", False),
        "comparison_end_to_end_perfect": result.get(
            "comparison_end_to_end_perfect", False),
        "open_product_findings_authorized": open_findings_authorized,
        "post_result_write_revalidation": postwrite_current,
        "post_result_write_identities": postwrite_detail,
    }
    _atomic_write_text(path, json.dumps(
        decision, indent=2, ensure_ascii=False) + "\n")
    return decision


def _parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--summary-xlsx-root", type=Path, default=DEFAULT_SUMMARY_XLSX_ROOT)
    parser.add_argument(
        "--summary-pdf-root", type=Path, default=DEFAULT_SUMMARY_PDF_ROOT)
    parser.add_argument("--tsn-raw", type=Path, default=DEFAULT_TSN_RAW)
    parser.add_argument("--tsn-xlsx", type=Path, default=DEFAULT_TSN_XLSX)
    parser.add_argument("--stage6-result", type=Path, default=DEFAULT_STAGE6_RESULT)
    parser.add_argument(
        "--stage6-acceptance", type=Path, default=DEFAULT_STAGE6_ACCEPTANCE)
    parser.add_argument(
        "--cross-format-result", type=Path, default=DEFAULT_CROSS_FORMAT_RESULT)
    parser.add_argument(
        "--tsnr-reference", type=Path, default=DEFAULT_TSNR_REFERENCE)
    parser.add_argument("--work-root", type=Path, default=DEFAULT_WORK_ROOT)
    parser.add_argument(
        "--allow-open-findings", action="store_true",
        help=(
            "accept the complete source/current-product oracle while the exact "
            "documented raw-conservation/fail-closed product findings remain open"))
    return parser.parse_args(argv)


def _emit(payload: dict[str, object]) -> None:
    sys.stdout.buffer.write((json.dumps(
        payload, ensure_ascii=False, separators=(",", ":")) + "\n").encode("utf-8"))


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(sys.argv[1:] if argv is None else argv)
    acceptance_path = args.output.with_suffix(args.output.suffix + ".acceptance.json")
    rejection_path = args.output.with_suffix(args.output.suffix + ".rejection.json")
    _unlink_if_present(acceptance_path)
    _unlink_if_present(rejection_path)
    try:
        result = run(args)
    except Exception as exc:
        failure = {
            "schema_version": 1,
            "audit": (
                "Stage 8 Intersection Summary authoritative TSMIS-vs-TSN base "
                "comparison oracle"),
            "error": {"type": type(exc).__name__, "message": str(exc)},
            "source_truth_exact": False,
            "production_value_projection_exact": False,
            "production_comparison_semantics_exact": False,
            "normalized_source_full_conservation": False,
            "stage8_base_oracle_complete": False,
            "comparison_end_to_end_perfect": False,
        }
        _atomic_write_text(args.output, json.dumps(
            failure, indent=2, ensure_ascii=False) + "\n")
        decision = _write_decision(
            rejection_path, args.output, failure, accepted=False,
            reason="oracle_execution_failed", postwrite_current=False,
            postwrite_detail={}, open_findings_authorized=False)
        _emit({
            "accepted": False, "reason": decision["reason"],
            "output": str(args.output), "rejection": str(rejection_path),
            "error": failure["error"],
        })
        return 2

    prewrite_current, prewrite_detail = _publication_current(args, result)
    result["publication_revalidation"] = {
        "after_complete_result_build": True,
        "before_result_write_all_current": prewrite_current,
        "before_result_write_identities": prewrite_detail,
    }
    result["audit_invariants"]["publication_inputs_current_before_write"] = (
        prewrite_current)
    result["stage8_base_oracle_complete"] = all(
        result["audit_invariants"].values())
    _atomic_write_text(args.output, json.dumps(
        result, indent=2, ensure_ascii=False) + "\n")

    postwrite_current, postwrite_detail = _publication_current(args, result)
    if not postwrite_current:
        result["publication_revalidation"]["post_result_write_all_current"] = False
        result["publication_revalidation"]["post_result_write_identities"] = (
            postwrite_detail)
        result["audit_invariants"]["publication_inputs_current_after_write"] = False
        result["stage8_base_oracle_complete"] = False
        _atomic_write_text(args.output, json.dumps(
            result, indent=2, ensure_ascii=False) + "\n")
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False,
            reason="post_result_write_revalidation_failed",
            postwrite_current=False, postwrite_detail=postwrite_detail,
            open_findings_authorized=False)
        _emit({"accepted": False, "reason": decision["reason"],
               "rejection": str(rejection_path)})
        return 2

    open_findings = bool(result["findings"]["product_red"])
    accepted = bool(
        result["source_truth_exact"]
        and result["production_value_projection_exact"]
        and result["production_comparison_semantics_exact"]
        and result["stage8_base_oracle_complete"]
        and postwrite_current
        and (not open_findings or args.allow_open_findings))
    if not accepted:
        reason = (
            "open_product_findings_not_authorized"
            if (result["stage8_base_oracle_complete"] and open_findings
                and not args.allow_open_findings)
            else "audit_or_projection_incomplete")
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False, reason=reason,
            postwrite_current=postwrite_current,
            postwrite_detail=postwrite_detail,
            open_findings_authorized=False)
        _emit({
            "accepted": False, "reason": reason, "output": str(args.output),
            "result_bytes": decision["result_bytes"],
            "result_sha256": decision["result_sha256"],
            "rejection": str(rejection_path),
        })
        return 1 if reason == "open_product_findings_not_authorized" else 2

    decision = _write_decision(
        acceptance_path, args.output, result, accepted=True,
        reason="oracle_complete_with_documented_product_findings",
        postwrite_current=postwrite_current, postwrite_detail=postwrite_detail,
        open_findings_authorized=bool(args.allow_open_findings and open_findings))
    acceptance_identity = _file_identity(acceptance_path)
    _emit({
        "accepted": True, "output": str(args.output),
        "result_bytes": decision["result_bytes"],
        "result_sha256": decision["result_sha256"],
        "acceptance": str(acceptance_path),
        "acceptance_bytes": acceptance_identity["bytes"],
        "acceptance_sha256": acceptance_identity["sha256"],
        "source_truth_exact": result["source_truth_exact"],
        "production_value_projection_exact": result[
            "production_value_projection_exact"],
        "production_comparison_semantics_exact": result[
            "production_comparison_semantics_exact"],
        "normalized_source_full_conservation": result[
            "normalized_source_full_conservation"],
        "stage8_base_oracle_complete": result["stage8_base_oracle_complete"],
        "product_findings": len(result["findings"]["product_red"]),
    })
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
