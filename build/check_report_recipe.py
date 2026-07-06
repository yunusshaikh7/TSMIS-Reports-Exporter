"""R3 (v0.19.0): the "add a report family" recipe is PROVEN, not just documented.

The user's next feature is the Highway Detail / Highway Summary comparison
(reserved-but-DISABLED stable ids 8/9 since v0.18.1). This check dry-runs the
docs/reports.md recipe end-to-end against the v0.19.0 substrate WITHOUT enabling
anything, so the moment the feature starts, every extension point is known-good:

  1. the reserved groundwork is intact (ids at frozen positions 8/9, app-wide
     disabled, greyed in the picker, absent from consolidate/compare);
  2. a hypothetical Highway Detail family REGISTERS through the catalog: the
     display views group it under the existing Highway family, the consolidator
     map and TSN entries accept it (in-memory patch — product tuples untouched);
  3. a ~20-line stub comparator built on compare_tsn_common.run_files_compare
     RUNS end-to-end (fixtures -> a real comparison workbook); and
  4. pdf_table_lib's writer produces the TSMIS-format route workbook a new
     PDF-sourced report would consolidate from.

Stdlib + openpyxl; no browser, no network. Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_report_recipe.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from openpyxl import Workbook, load_workbook

import batch_manifest
import compare_tsn_common as ctc
import pdf_table_lib
import report_catalog as cat
import reports
from compare_core import CompareSchema

_fail = []


def check(name, cond, detail=""):
    if cond:
        print(f"  ok: {name}")
    else:
        print(f"FAIL: {name}" + (f"\n      {detail}" if detail else ""))
        _fail.append(name)


def test_reserved_groundwork():
    print("the reserved Highway Detail/Summary ids (8/9) are intact + disabled:")
    keys = [e.key for e in cat.EXPORT]
    check("stable ids at the frozen positions 8/9",
          keys[8] == "highway_detail" and keys[9] == "highway_summary", str(keys))
    check("batch manifest order matches (append-only contract)",
          list(batch_manifest._V017_EXPORT_ORDER[8:10])
          == ["highway_detail", "highway_summary"])
    check("both are app-wide DISABLED (greyed, rejected server-side)",
          {"highway_detail", "highway_summary"} <= set(reports.DISABLED_EXPORT_SUBDIRS))
    check("their stub specs refuse to save",
          _raises(lambda: cat.EXPORT[8].spec.save(None, None, None)))
    check("absent from consolidate/compare until the feature lands",
          not any("highway_detail" in e.key or "highway_summary" in e.key
                  for e in cat.CONSOLIDATE + cat.COMPARE))


def _raises(fn):
    try:
        fn()
        return False
    except Exception:
        return True


class _Patch:
    def __init__(self, obj, name, value):
        self.obj, self.name, self.value = obj, name, value

    def __enter__(self):
        self.old = getattr(self.obj, self.name)
        setattr(self.obj, self.name, self.value)

    def __exit__(self, *a):
        setattr(self.obj, self.name, self.old)


def test_catalog_registration():
    print("a hypothetical Highway Detail family registers through the catalog:")

    class _Mod:                                       # a stub consolidator module
        REPORT_NAME = "Highway Detail"
        INPUT_FMT = "Excel"
        subdir = "highway_detail"

    new_cons = cat.CONSOLIDATE + (
        cat.ConsolidateEntry("cons:highway_detail", "Highway Detail", _Mod),)
    new_tsn = cat.TSN + (
        cat.TsnEntry("highway_detail", "TSN Highway Detail", "*.xlsx",
                     "statewide_xlsx", "tsn_highway_detail_normalized.xlsx",
                     "tsn_load_highway_detail:build_into", normalization_version=1),)
    new_auto = tuple(cat._AUTO_CONSOLIDATOR) + (("highway_detail", _Mod),)
    with _Patch(cat, "CONSOLIDATE", new_cons), _Patch(cat, "TSN", new_tsn), \
         _Patch(cat, "_AUTO_CONSOLIDATOR", new_auto):
        order, meta = cat.consolidate_display()
        check("the new consolidator appears in the display view",
              "cons:highway_detail" in order)
        check("...grouped under the EXISTING Highway family (W2 metadata)",
              meta["cons:highway_detail"] == cat.export_display()["highway_detail"],
              str(meta.get("cons:highway_detail")))
        check("...contiguous with its family in picker order",
              cat._picker_family("cons:highway_detail") == "highway_detail")
        check("consolidator_by_subdir maps the new subdir",
              cat.consolidator_by_subdir()["highway_detail"] is _Mod)
        entries = {e.subdir: e for e in cat.tsn_entries()}
        check("tsn_entries carries the new report + its normalization version",
              entries["highway_detail"].normalization_version == 1)


def test_stub_comparator_runs():
    print("a run_files_compare stub comparator runs end-to-end (the R1 substrate):")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_recipe_"))

    def _wb(path, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Highway Detail"
        ws.append(["Route", "PM", "Value"])
        for r in rows:
            ws.append(r)
        wb.save(path)
        wb.close()

    a, b = tmp / "tsmis.xlsx", tmp / "tsn.xlsx"
    _wb(a, [["001", "1.000", "X"], ["001", "2.000", "Y"]])
    _wb(b, [["001", "1.000", "X"], ["001", "2.000", "Z"]])

    schema = CompareSchema(report_name="Highway Detail",
                           header=["PM", "Value"], side_a="TSMIS", side_b="TSN",
                           id_noun="location", id_noun_plural="locations",
                           pair_noun="postmile", sides_noun="systems")

    def loader(pa, pb):
        def rows(p):
            return ctc.load_consolidated_rows(
                p, "Highway Detail",
                missing_sheet_hint="pick the consolidated Highway Detail workbook.",
                bad_header_msg="isn't a CONSOLIDATED Highway Detail workbook.")[0]
        return rows(pa), rows(pb), None

    res = ctc.run_files_compare(
        schema, a, b, tmp / "out.xlsx",
        banner="Highway Detail Comparison — TSMIS vs TSN",
        has_route=True, loader=loader,
        confirm_overwrite=lambda p: True, mode="values")
    check("the stub comparator produces a comparison workbook",
          res.status == "ok" and Path(res.output_path).is_file(),
          getattr(res, "message", ""))
    wb = load_workbook(res.output_path, read_only=True)
    try:
        check("...with the engine's sheet set",
              all(s in wb.sheetnames for s in ("Summary", "Comparison")),
              str(wb.sheetnames))
    finally:
        wb.close()
    check("...and the family filename helper works",
          ctc.suggest_route_name("x route 7.xlsx", "Highway_Detail",
                                 "TSMIS_vs_TSN_HighwayDetail")
          .startswith("TSMIS_vs_TSN_HighwayDetail_Route7_Comparison"))


def test_pdf_writer_recipe():
    print("pdf_table_lib writes the TSMIS-format route workbook a new report needs:")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_recipe_pdf_"))
    out = tmp / "route.xlsx"
    header = ["PM", "Description", "Value"]
    pdf_table_lib.write_route_workbook(
        [["1.000", "=INJECTED()", "A"]], out, sheet_name="Highway Detail",
        header=header)
    wb = load_workbook(out)
    try:
        ws = wb["Highway Detail"]
        check("header + freeze panes + widths applied",
              [c.value for c in ws[1]] == header and ws.freeze_panes == "A2")
        check("formula-injection neutralized in the data rows",
              ws.cell(row=2, column=2).data_type == "s")
    finally:
        wb.close()


def main():
    test_reserved_groundwork()
    test_catalog_registration()
    test_stub_comparator_runs()
    test_pdf_writer_recipe()
    if _fail:
        print(f"\n{len(_fail)} check(s) FAILED")
        return 1
    print("\nall good")
    return 0


if __name__ == "__main__":
    sys.exit(main())
