"""Golden checks for the cross-environment loader's per-route universe integrity
(CMP-AUD-027 + CMP-AUD-030 + CMP-AUD-031 in scripts/compare_env.py::_load_xlsx_side).

The flat per-route XLSX loader keys every side by the route pulled from each
"<report>_route_<token>.xlsx" export name. Three silent-corruption holes lived in
that mapping:

  CMP-AUD-027  a valid-header file contributing ZERO data rows appended no
               [route, …] row, so its route silently VANISHED from coverage —
               the comparison could then certify a clean match while a whole
               route present on one side was invisible. The statewide census
               (756 real per-route exports; min 1 data row) found NO header-only
               file, so a data-less export is anomalous: it is now disclosed
               LOUDLY as an incomplete input naming the route, never dropped.


  CMP-AUD-031  the route token was used RAW (the uppercased capture, or an
               arbitrary file stem when no "_route_" pattern matched), never run
               through the same zero-pad normalizer the Ramp Summary path uses.
               So "route_1.xlsx" and "route_001.xlsx" keyed as two DIFFERENT
               routes (a route split into two one-sided rows), and a canonical
               workbook named "totally_unrelated.xlsx" was promoted to a bogus
               route "TOTALLY_UNRELATED" and cleanly matched.

  CMP-AUD-030  no seen-route set existed, so two files that resolve to the SAME
               route on one side were silently concatenated (a stale copy or a
               split export doubling coverage) with no input diagnostic.

The fix normalizes the token, requires the "..._route_<n>" export naming
contract (a non-route file is skipped LOUDLY, never promoted to a route), and
rejects a duplicate route on a side into the existing `skipped` incompleteness
channel — so neither hole can masquerade as a clean match. Canonical distinct
routes (including "005" vs "005S", which must stay separate) are unaffected.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_env_route_universe.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env as env
from events import Events
from openpyxl import Workbook

_SHEET = "Highway Locations"
_HEADER = ["County", "City", "R", "PM", "X", "Description"]


def _make_side(files):
    """files: [(filename, [data_row, ...]), ...]. Build a highway_sequence side
    folder with one XLSX per entry (all sharing _HEADER) and return its parent."""
    base = Path(tempfile.mkdtemp())
    d = base / "highway_sequence"
    d.mkdir(parents=True)
    for name, data in files:
        wb = Workbook()
        ws = wb.active
        ws.title = _SHEET
        ws.append(_HEADER)
        for r in data:
            ws.append(r)
        wb.save(d / name)
    return base


def _load(base):
    return env._load_xlsx_side(base, "X", "highway_sequence", _SHEET,
                               "Highway Sequence", Events())


def test_031_route_token_normalized():
    """route_1 must resolve to the normalized route 001 (not the raw "1"), so a
    cross-side pairing can't split one route into two one-sided rows."""
    base = _make_side([("hs_route_1.xlsx", [["ORA", "", "R", "1.0", "", "A"]])])
    rows, _header, skipped = _load(base)
    assert rows, "the route_1 export must contribute a row"
    assert rows[0][0] == "001", (
        "route_1 must normalize to 001 (was the raw token)", rows[0][0])
    assert not skipped, skipped


def test_031_non_route_name_rejected():
    """A workbook without the ..._route_<n> contract must NOT be promoted to a
    route identity from its stem — it is skipped LOUDLY. The real route export
    beside it still contributes."""
    base = _make_side([
        ("hs_route_005.xlsx", [["ORA", "", "R", "5.0", "", "V"]]),
        ("totally_unrelated.xlsx", [["ORA", "", "R", "1.0", "", "A"]]),
    ])
    rows, _header, skipped = _load(base)
    routes = {r[0] for r in rows}
    assert routes == {"005"}, (
        "only the real per-route export may contribute a route", routes)
    assert any("totally_unrelated" in s for s in skipped), (
        "the non-route file must be disclosed as skipped", skipped)


def test_030_duplicate_route_flagged():
    """Two files that resolve to the same route on one side must not silently
    concatenate — the duplicate is skipped into the incompleteness channel."""
    base = _make_side([
        ("a_route_001.xlsx", [["ORA", "", "R", "1.0", "", "PM1"]]),
        ("b_route_001.xlsx", [["ORA", "", "R", "2.0", "", "PM2"]]),
    ])
    rows, _header, skipped = _load(base)
    assert len(rows) == 1, (
        "duplicate-route files must not concatenate their rows", len(rows))
    assert any("duplicate" in s.lower() and "001" in s for s in skipped), (
        "the duplicate route must be disclosed as skipped", skipped)


def test_029_owner_lock_ignored():
    """An Office owner-lock stub (~$...) beside a real per-route export must be
    ignored entirely — not counted as a member and skipped-as-incomplete.
    Merely having the workbook open in Excel cannot turn identical exports into
    an incomplete comparison (CMP-AUD-029)."""
    base = _make_side([("hs_route_001.xlsx", [["ORA", "", "R", "1.0", "", "A"]])])
    lock = base / "highway_sequence" / "~$hs_route_001.xlsx"
    lock.write_bytes(b"\x00\x02lockstub-not-a-workbook")  # an Excel lock stub
    rows, _header, skipped = _load(base)
    assert {r[0] for r in rows} == {"001"}, [r[0] for r in rows]
    assert not skipped, (
        "the owner-lock stub must be ignored, not skipped as incomplete", skipped)


def test_027_header_only_route_disclosed():
    """A valid-header file with ZERO data rows must not silently vanish: its
    route is disclosed in the incompleteness channel (naming it), while a real
    per-route file beside it still contributes its row."""
    base = _make_side([
        ("hs_route_001.xlsx", [["ORA", "", "R", "1.0", "", "REAL"]]),
        ("hs_route_002.xlsx", []),                      # header only — no data rows
    ])
    rows, _header, skipped = _load(base)
    routes = {r[0] for r in rows}
    assert routes == {"001"}, ("only the route with data contributes a row", routes)
    assert any("002" in s and "no data rows" in s for s in skipped), (
        "the header-only route must be disclosed as incomplete, not dropped", skipped)


def test_027_header_only_sole_file_errors():
    """A side whose ONLY file is header-only has no data to compare: it errors
    LOUDLY (the existing whole-empty-side guard) — never a silent clean side."""
    base = _make_side([("hs_route_002.xlsx", [])])
    try:
        _load(base)
    except ValueError as e:
        assert "No readable" in str(e), str(e)
    else:
        raise AssertionError("a sole header-only file must raise, not return empty")


def test_027_header_only_end_to_end_incomplete():
    """The finding's fixture end to end: side A = route 001 (real) + route 002
    (header-only); side B = route 001 only. Pre-fix the comparison certified a
    clean match with route 002 erased; now it is flagged INCOMPLETE and route
    002 is named, even though the shared route 001 matches."""
    a = _make_side([
        ("hs_route_001.xlsx", [["ORA", "", "R", "1.0", "", "SAME"]]),
        ("hs_route_002.xlsx", []),                      # header only
    ])
    b = _make_side([("hs_route_001.xlsx", [["ORA", "", "R", "1.0", "", "SAME"]])])
    adapter = env.EnvCompare("hs_test", "Highway Sequence", "highway_sequence",
                             sheet_name=_SHEET, key_col="PM")
    out = Path(tempfile.mkdtemp()) / "cmp.xlsx"
    result = adapter.compare_folders(a, b, out, events=Events(),
                                     confirm_overwrite=lambda _p: True,
                                     mode="formulas")
    assert result.status == "ok", (result.status, result.message)
    assert result.completion == "partial", (
        "a header-only route on a side must make the comparison INCOMPLETE, "
        "not a clean match", result.completion)
    warns = " ".join(result.comparison_outcome.warnings)
    assert "002" in warns and "no data rows" in warns, (
        "route 002's absence-of-data must be named in the coverage warnings", warns)


def test_030_031_canonical_side_unchanged():
    """Positive control: distinct canonical routes — including 005 vs 005S,
    which must stay SEPARATE — all contribute with zero skips."""
    base = _make_side([
        ("hs_route_001.xlsx", [["ORA", "", "R", "1.0", "", "A"]]),
        ("hs_route_005.xlsx", [["ORA", "", "R", "5.0", "", "B"]]),
        ("hs_route_005S.xlsx", [["ORA", "", "R", "5.5", "", "C"]]),
    ])
    rows, _header, skipped = _load(base)
    routes = sorted(r[0] for r in rows)
    assert routes == ["001", "005", "005S"], (
        "005 and 005S must stay distinct and all contribute", routes)
    assert not skipped, skipped


def test_044_data_beyond_header_disclosed():
    """CMP-AUD-044: a per-route export with a TRAILING BLANK header column that
    still carries data must not have that column silently sliced off (r[:n]) and
    dropped from the comparison. The loader refuses the file LOUDLY (an
    incompleteness skip) naming it, while a clean file beside it contributes."""
    base = Path(tempfile.mkdtemp())
    d = base / "highway_sequence"
    d.mkdir(parents=True)
    clean = Workbook()
    ws = clean.active
    ws.title = _SHEET
    ws.append(_HEADER)
    ws.append(["ORA", "", "R", "1.0", "", "CLEAN"])
    clean.save(d / "hs_route_001.xlsx")
    bad = Workbook()
    ws = bad.active
    ws.title = _SHEET
    ws.append(_HEADER + [""])                                   # trailing blank header
    ws.append(["ORA", "", "R", "2.0", "", "DESC", "DROPPED"])   # data beyond the header
    bad.save(d / "hs_route_002.xlsx")
    rows, _header, skipped = _load(base)
    assert {r[0] for r in rows} == {"001"}, (
        "only the clean route contributes; the overflow file must be refused",
        [r[0] for r in rows])
    assert any("002" in s and "beyond" in s.lower() for s in skipped), (
        "the trailing-blank-header data file must be disclosed as skipped", skipped)


def main():
    test_031_route_token_normalized()
    test_031_non_route_name_rejected()
    test_030_duplicate_route_flagged()
    test_029_owner_lock_ignored()
    test_027_header_only_route_disclosed()
    test_027_header_only_sole_file_errors()
    test_027_header_only_end_to_end_incomplete()
    test_030_031_canonical_side_unchanged()
    test_044_data_beyond_header_disclosed()
    print("OK  cross-env route universe: tokens zero-pad-normalized, the "
          "..._route_<n> naming contract is required (no promoted stems), "
          "duplicate routes are disclosed as incomplete, header-only routes are "
          "disclosed (never silently dropped), and distinct canonical routes "
          "(005 vs 005S) are unaffected.")


if __name__ == "__main__":
    main()
