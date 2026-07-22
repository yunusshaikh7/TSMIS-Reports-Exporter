"""Golden check for the published-comparison spine (CMP-AUD-208 / 209 / 108).

Evidence used to prove itself by RE-RUNNING the loaders that produced the
comparison — a consistently wrong loader passed that test twice. These checks
lock the replacement: a decoder that reads what the workbook PUBLISHED (its
hidden per-cell E/D/N/U state masks, its anchored Status/Diffs contract, its
opaque row tokens), an EXHAUSTIVE hash-bound ledger built before any sample is
drawn, and a reconciler that lets the published cell — not a second execution —
decide whether an image may be rendered.

Every fixture is a REAL comparison workbook built by ``compare_core.run_compare``
in the matrix's own ``mode="values"``, so the decoder is exercised against the
bytes the product actually ships. The mutation matrix then corrupts one thing at
a time and requires refusal: a decoder that cannot authenticate the published
cells must not be allowed to claim it verified them.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_published_comparison.py
"""
import os
import shutil
import sys
import tempfile
from collections import Counter
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import published_comparison as pc
import visual_evidence as ve
from compare_core import CompareSchema, published_key_text, run_compare
from openpyxl import load_workbook

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def raises(fn, fragment=""):
    try:
        fn()
    except pc.PublishedComparisonError as e:
        return fragment.lower() in str(e).lower()
    except Exception:                            # noqa: BLE001
        return False
    return False


# --------------------------------------------------------------------------- #
# fixture: a real values comparison over a compact, fully-classified corpus
# --------------------------------------------------------------------------- #
HEADER = ["PM", "County", "Note", "FT", "Description"]
SCHEMA = CompareSchema(report_name="Check", header=HEADER, key_field=0,
                       context_fields=("Note",))
FIELDS = ["County", "Note", "FT", "Description"]      # display order, key first


def row(route, pm, county="ALA", note="n", ft="1", desc="D"):
    return [route, pm, county, note, ft, desc]


# side A / side B: one identical pair, one differing pair, one context-only
# difference, one one-sided row per side, and a duplicate group whose ONLY
# differences live inside it (the CMP-AUD-108 shape).
ROWS_A = [
    row("001", "1.000"),
    row("001", "2.000", desc="ALPHA"),
    row("001", "3.000", note="left"),
    row("001", "4.000"),                          # A-only
    row("002", "9.000", desc="DUP-A1"),           # duplicate group
    row("002", "9.000", desc="DUP-A2"),
]
ROWS_B = [
    row("001", "1.000"),
    row("001", "2.000", desc="BETA"),
    row("001", "3.000", note="right"),
    row("001", "5.000"),                          # B-only
    row("002", "9.000", desc="DUP-B1"),
    row("002", "9.000", desc="DUP-B2"),
]

tmp = Path(tempfile.mkdtemp(prefix="check_pubcmp_"))
values = tmp / "cmp.xlsx"
result = run_compare(SCHEMA, ROWS_A, ROWS_B, True, values, mode="values",
                     confirm_overwrite=lambda _p: True)

print("fixture")
check("the fixture comparison published", result.status == "ok" and values.is_file())

# --------------------------------------------------------------------------- #
print("decoding the published cells (CMP-AUD-208)")
published = pc.read(values)
check("the displayed columns decode in publication order",
      list(published.fields) == FIELDS)
check("the state-mask version is the engine's own",
      published.state_version.startswith("CMP_E") and "_STATE_V" in
      published.state_version)
check("the row-token version is the engine's own",
      "_KEY_V" in published.token_version)
check("both side labels come from the workbook's own source-row columns",
      published.side_labels == ("TSMIS", "TSN"))
check("every published row decoded", len(published.rows) == 7)

identical = published.row_at("001", "1.000")
differing = published.row_at("001", "2.000")
contextual = published.row_at("001", "3.000")
a_only = published.row_at("001", "4.000")
b_only = published.row_at("001", "5.000")
check("an identical matched row is addressable", identical is not None)
check("a differing matched row is addressable", differing is not None)
check("the fully-identical row is all-equal except its context cell",
      identical is not None
      and identical.mask == "ENEE" and identical.diffs == 0)
check("the differing row marks exactly its Description cell",
      differing is not None and differing.mask == "ENED"
      and differing.diffs == 1)
check("a context-only difference is N, never counted",
      contextual is not None and contextual.mask == "ENEE"
      and contextual.diffs == 0)
check("one-sided rows carry an all-U mask and no Diffs",
      a_only is not None and b_only is not None
      and set(a_only.mask) == {"U"} and a_only.diffs == 0
      and set(b_only.mask) == {"U"})
dup_rows = [r for r in published.rows if r.route == "002"]
check("the duplicate group publishes both occurrences",
      len(dup_rows) == 2
      and published.group_size("002", dup_rows[0].key) == 2)
check("the duplicate occurrences are numbered 1..n",
      sorted(r.occurrence for r in dup_rows) == [1, 2])
check("a duplicate-group row is not solo",
      all(not published.is_solo(r) for r in dup_rows))
check("a unique-key row is solo", published.is_solo(differing))

check("require_fields accepts the published columns",
      published.require_fields(FIELDS) is None)
check("require_fields refuses a column the comparison does not publish",
      raises(lambda: published.require_fields(["County", "Nope"]), "Nope"))
check("position_of refuses an unknown column",
      raises(lambda: published.position_of("Nope"), "no column"))

# --------------------------------------------------------------------------- #
print("the exhaustive ledger, built before any sample (CMP-AUD-209)")
ledger = published.ledger()
check("the ledger counts every published row", ledger.data_rows == 7)
check("matched and one-sided rows partition the universe",
      ledger.matched_rows + ledger.one_sided_rows == ledger.data_rows
      and ledger.one_sided_rows == 2)
check("one-sided rows are named per side",
      dict(ledger.one_sided_by_status) == {"TSMIS only": 1, "TSN only": 1})
check("the repeated-key group is inventoried",
      ledger.duplicate_groups == 1 and ledger.duplicate_member_rows == 2)
check("counted differences equal the sum of the per-column counts",
      ledger.difference_cells
      == sum(e.differences for e in ledger.fields))
check("Description's differences split into unique-row and duplicate-group",
      ledger.differences("Description") == 3
      and ledger.solo_differences("Description") == 1
      and ledger.duplicate_differences("Description") == 2)
check("a context column is inventoried but never counted",
      ledger.differences("Note") == 0
      and ledger.for_field("Note").context_cells == 5)
check("one-sided cells are inventoried, not dropped",
      ledger.one_sided_cells == 2 * len(FIELDS))
check("fields_with_differences names Description only",
      ledger.fields_with_differences() == ("Description",))

digest = ledger.digest()
check("the ledger digest is stable across reads",
      digest == published.ledger().digest() and len(digest) == 64)
check("the ledger digest is hash-bound to its own contents",
      pc.ComparisonLedger(**{**{k: v for k, v in
                               vars(ledger).items() if not k.startswith("_")},
                             "difference_cells": 999},
                          _by_field={}).digest() != digest)

# --------------------------------------------------------------------------- #
print("persisted source rows through the opaque token (CMP-AUD-208)")
resolved = published.source_rows([differing.token])
check("a matched row resolves BOTH persisted source rows",
      set(resolved.get(differing.token, {})) == {"TSMIS", "TSN"})
check("an unknown token resolves to nothing",
      published.source_rows(["__nope__"]) == {})
check("no tokens asked for means no scan", published.source_rows([]) == {})
check("row tokens are unique across the sheet",
      len({r.token for r in published.rows}) == len(published.rows))

# --------------------------------------------------------------------------- #
print("reconciliation: the published cell decides (CMP-AUD-208)")
pub_key = differing.key


def candidate(**over):
    base = dict(route="001", key="2.000", field="Description",
                va="ALPHA", vb="BETA", dist="", cnty="",
                pub_key=pub_key, display=differing.value(
                    published.position_of("Description")))
    base.update(over)
    return base


kept, rejected = ve._reconcile(published, {"Description": [candidate()]})
check("a candidate that matches the published cell is kept",
      len(kept.get("Description", [])) == 1 and not rejected)
entry = kept["Description"][0]
check("the kept candidate names its published row/occurrence/state",
      entry["published_row"] == differing.excel_row
      and entry["published_occurrence"] == 1
      and entry["published_state"] == "D"
      and entry["published_token"] == differing.token)

_, rejected = ve._reconcile(
    published, {"Description": [candidate(pub_key="nonexistent")]})
check("a candidate with no published row is refused",
      rejected["Description"][ve._REJECT_NO_ROW] == 1)
_, rejected = ve._reconcile(
    published, {"Description": [candidate(display="SOMETHING ELSE")]})
check("a candidate whose text disagrees with the published cell is refused",
      rejected["Description"][ve._REJECT_TEXT] == 1)
_, rejected = ve._reconcile(
    published, {"County": [candidate(field="County",
                                     display=identical.value(0))]})
check("a candidate on a cell the comparison did not count is refused",
      rejected["County"][ve._REJECT_NOT_COUNTED] == 1)
dup = [r for r in dup_rows if r.occurrence == 1][0]
_, rejected = ve._reconcile(
    published, {"Description": [candidate(route="002", pub_key=dup.key,
                                          display=dup.value(
                                              published.position_of(
                                                  "Description")))]})
check("a candidate inside a repeated-key group is refused as ambiguous",
      rejected["Description"][ve._REJECT_NOT_SOLO] == 1)
_, rejected = ve._reconcile(published, {"Nope": [candidate(field="Nope")]})
check("a candidate on an unpublished column is refused, not raised",
      rejected["Nope"][ve._REJECT_NO_ROW] == 1)

# --------------------------------------------------------------------------- #
print("duplicate-only differences are NAMED, never a false zero (CMP-AUD-108)")
dup_only = pc.ComparisonLedger(
    fields=(pc.FieldLedger(field="Description", position=3, differences=2,
                           solo_differences=0, duplicate_differences=2),),
    data_rows=2, matched_rows=2, one_sided_rows=0, one_sided_by_status=(),
    duplicate_groups=1, duplicate_member_rows=2, difference_cells=2,
    context_cells=0, equal_cells=0, one_sided_cells=0,
    _by_field={})
reason = ve._unrenderable_reason(dup_only.fields[0], None)
check("a duplicate-only column reports a named miss",
      reason is not None and "repeated-key" in reason and "2" in reason)
check("the named miss never claims there is no difference",
      "no differing" not in (reason or "").lower())
check("a column refused by the published cells names that instead",
      "refused" in (ve._unrenderable_reason(
          pc.FieldLedger(field="FT", position=2, differences=1,
                         solo_differences=1),
          Counter({ve._REJECT_TEXT: 4})) or ""))
check("a renderable column with candidates has no pre-render miss",
      ve._unrenderable_reason(
          pc.FieldLedger(field="FT", position=2, differences=1,
                         solo_differences=1), None) is None)

# --------------------------------------------------------------------------- #
print("published-key addressing uses the engine's own identity")
check("published_key_text reproduces the published key cell",
      published_key_text(SCHEMA, ROWS_A[1]) == differing.key)
check("the ledger rows carry every displayed column",
      [r[0] for r in ve._ledger_rows(ledger, FIELDS, Counter())] == FIELDS)
check("the ledger rows report the sampled count per column",
      dict((r[0], r[7]) for r in ve._ledger_rows(
          ledger, FIELDS, Counter({"Description": 2})))["Description"] == 2)


# --------------------------------------------------------------------------- #
print("the mutation matrix — an unauthenticatable workbook is refused")


def mutate(tag, fn):
    """Copy the fixture, apply one corruption, and return the new path."""
    path = tmp / f"mut_{tag}.xlsx"
    shutil.copy2(values, path)
    wb = load_workbook(path)
    fn(wb["Comparison"], wb)
    wb.save(path)
    return path


def header_index(sheet, predicate):
    for cell in sheet[1]:
        if predicate(cell.value):
            return cell.column
    raise AssertionError("header column not found")


def state_col(sheet):
    return header_index(sheet, lambda v: isinstance(v, str)
                        and "_STATE_V" in v)


def token_col(sheet):
    return header_index(sheet, lambda v: isinstance(v, str)
                        and v.endswith("_TOKEN"))


def diffs_col(sheet):
    return header_index(sheet, lambda v: v == "Diffs")


def status_col(sheet):
    return header_index(sheet, lambda v: v == "Status")


cases = [
    ("short_mask", "state codes",
     lambda ws, wb: ws.cell(row=2, column=state_col(ws), value="EN")),
    ("bad_code", "unknown state code",
     lambda ws, wb: ws.cell(row=2, column=state_col(ws), value="ENEX")),
    ("diffs_lie", "state mask holds",
     lambda ws, wb: ws.cell(row=2, column=diffs_col(ws), value=4)),
    ("matched_u", "one-sided state codes",
     lambda ws, wb: ws.cell(row=2, column=state_col(ws), value="ENEU")),
    ("onesided_diffs", "carries",
     lambda ws, wb: ws.cell(
         row=[c.row for c in ws["F"] if c.value in ("TSMIS only", "TSN only")][0],
         column=diffs_col(ws), value=2)),
    ("onesided_mask", "not entirely one-sided",
     lambda ws, wb: ws.cell(
         row=[c.row for c in ws["F"] if c.value in ("TSMIS only", "TSN only")][0],
         column=state_col(ws), value="UUUE")),
    ("dup_token", "repeats row token",
     lambda ws, wb: ws.cell(row=3, column=token_col(ws),
                            value=ws.cell(row=2, column=token_col(ws)).value)),
    ("no_status", "Status",
     lambda ws, wb: ws.cell(row=1, column=status_col(ws), value="Statuz")),
    ("dup_status", "exactly one is required",
     lambda ws, wb: ws.cell(row=1, column=1, value="Status")),
    ("no_state", "no state-mask chunks",
     lambda ws, wb: ws.cell(row=1, column=state_col(ws), value="Notes")),
    ("no_token", "exactly one hidden row-token",
     lambda ws, wb: ws.cell(row=1, column=token_col(ws), value="helper")),
    ("bad_occ", "non-integer occurrence",
     lambda ws, wb: ws.cell(row=2, column=3, value="first")),
    # NOTE: openpyxl treats `cell(..., value=None)` as "no value supplied" and
    # leaves the cell alone — a blanking mutation must assign .value directly.
    ("blank_status", "no Status",
     lambda ws, wb: setattr(ws.cell(row=2, column=status_col(ws)),
                            "value", None)),
    ("no_sheet", "Comparison",
     lambda ws, wb: setattr(ws, "title", "Comparisons")),
]
for tag, fragment, fn in cases:
    path = mutate(tag, fn)
    check(f"refused: {tag}", raises(lambda p=path: pc.read(p), fragment))

check("the untouched fixture still reads after the mutation sweep",
      pc.read(values).ledger().digest() == digest)
check("expect_fields refuses a comparison of other columns",
      raises(lambda: pc.read(values, expect_fields=["A", "B"]),
             "not the ones this report compares"))
check("expect_fields accepts the published columns",
      len(pc.read(values, expect_fields=FIELDS).rows) == 7)

missing = tmp / "gone.xlsx"
check("a missing workbook is refused, never guessed",
      raises(lambda: pc.read(missing), "could not be opened"))

shutil.rmtree(tmp, ignore_errors=True)

print()
if _fail:
    print(f"FAILED {len(_fail)} check(s):")
    for name in _fail:
        print(f"  - {name}")
    sys.exit(1)
print("published-comparison spine: all checks passed")
