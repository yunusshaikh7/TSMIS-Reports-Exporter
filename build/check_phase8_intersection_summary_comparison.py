#!/usr/bin/env python3
"""Permanent adversarial gate for the Stage-8 Intersection Summary oracle."""

from __future__ import annotations

import copy
import json
from pathlib import Path
import sys
import tempfile
import zipfile

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

import phase8_intersection_summary_comparison as oracle


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def rejected(callable_, message: str) -> None:
    try:
        callable_()
    except oracle.AuditError:
        return
    raise AssertionError(message)


def _write_tsn(path: Path, rows: list[tuple[str, object]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Intersection Summary (TSN)"
    ws.append(["Category", "Count"])
    for row in rows:
        ws.append(row)
    workbook.save(path)
    workbook.close()


def _write_tsnr(path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws["A1"] = "Intersection Data Item"
    ws["H1"] = "Crosswalk (TSN-TSNR)"
    ws["A2"] = "Intersection_Geometry"
    ws["A3"] = "Intersection_Control"
    ws["B3"] = "\n".join((
        "F. Four-Way Flasher (Red on Mainline)",
        "G. Four-Way flasher (Red on All)",
        "J. Signals Pretimed – 2 phase",
        "K. Signals Pretimed – Multi-Phase",
        "L. Signals Semi-Traffic Actuated – 2 Phase",
        "M. Signals Semi-Traffic Actuated – Multi Phase",
        "N. Signals Full-Traffic Actuated – 2 Phase",
        "P. Signals Full-Traffic Actuated – Multi-Phase",
    ))
    ws["E3"] = "5. Signalized"
    ws["H3"] = "\n".join(
        f"TSN: {code} source = TSNR: 5. Signalized" for code in "JKLMNP")
    workbook.save(path)
    workbook.close()


def _write_zip(path: Path, core: bytes, sheet: bytes,
               extra: tuple[str, bytes] | None = None) -> None:
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("docProps/core.xml", core)
        archive.writestr("xl/worksheets/sheet1.xml", sheet)
        if extra is not None:
            archive.writestr(*extra)


def _valid_record() -> dict[str, int]:
    return dict(oracle.EXPECTED_TSMIS_AGGREGATE)


def _publication_fixture(*, audit_complete: bool) -> dict[str, object]:
    return {
        "schema_version": 1,
        "audit": "synthetic Stage-8 Intersection Summary decision fixture",
        "audit_invariants": {"synthetic_audit": audit_complete},
        "source_truth_exact": audit_complete,
        "production_value_projection_exact": audit_complete,
        "production_comparison_semantics_exact": audit_complete,
        "normalized_source_full_conservation": False,
        "stage8_base_oracle_complete": audit_complete,
        "comparison_end_to_end_perfect": False,
        "findings": {"product_red": [{"finding": "synthetic"}]},
    }


def _pdf_words() -> list[dict[str, object]]:
    words = []
    geometry = {
        "left": (66.0, 130.0, 16.0),
        "center": (257.0, 130.0, 14.0),
        "right": (448.0, 130.0, 18.0),
    }
    for column in ("left", "center", "right"):
        x1, top, step = geometry[column]
        for index, _cat in enumerate(oracle.PDF_COLUMNS[column]):
            words.append({
                "text": str(index), "x0": x1 - 10, "x1": x1,
                "top": top + index * step,
                "bottom": top + index * step + 8,
            })
    return words


def main() -> int:
    for label, value in (("bool", True), ("float", 1.0),
                         ("string", "1"), ("negative", -1)):
        rejected(lambda value=value: oracle._strict_count(value, label),
                 f"strict count accepted {label}")
    require(oracle._strict_count(0, "zero") == 0,
            "strict count rejected zero")

    require(oracle._route_from_name(
        Path("intersection_summary_route_008U.pdf"), ".pdf") == "008U",
        "suffix route identity was truncated")
    require(oracle._route_from_name(
        Path("intersection_summary_route_010S.xlsx"), ".xlsx") == "010S",
        "valid suffix XLSX identity was rejected")
    rejected(lambda: oracle._route_from_name(
        Path("intersection_summary_route_8.pdf"), ".pdf"),
        "unpadded route filename was accepted")

    valid_routes = ["001", "008U", "905"]
    valid_digest = oracle._route_lf_digest(valid_routes)
    exact = oracle._require_exact_route_universes(
        {"xlsx": valid_routes, "pdf": list(valid_routes)},
        expected_count=3, expected_digest=valid_digest)
    require(exact["all_exact"], "valid route universes did not pass")
    rejected(lambda: oracle._require_exact_route_universes(
        {"xlsx": ["001", "001", "905"], "pdf": valid_routes},
        expected_count=3, expected_digest=valid_digest),
        "duplicate route universe was accepted")
    rejected(lambda: oracle._require_exact_route_universes(
        {"xlsx": ["001", "008U"], "pdf": ["001", "008U"]},
        expected_count=3, expected_digest=valid_digest),
        "dropped route universe was accepted")
    rejected(lambda: oracle._require_exact_route_universes(
        {"xlsx": valid_routes, "pdf": ["001", "905", "008U"]},
        expected_count=3, expected_digest=valid_digest),
        "route-order mutation was accepted")
    rejected(lambda: oracle._require_exact_route_universes(
        {"xlsx": ["001", "170", "905"], "pdf": ["001", "170", "905"]},
        expected_count=3, expected_digest=valid_digest),
        "route-170 injection escaped the frozen route digest")
    rejected(lambda: oracle._require_exact_route_universes(
        {"xlsx": ["001", "008", "905"], "pdf": ["001", "008", "905"]},
        expected_count=3, expected_digest=valid_digest),
        "suffix collapse escaped the frozen route digest")

    valid_record = _valid_record()
    arithmetic = oracle._validate_record(valid_record, "valid")
    require(all(value == oracle.EXPECTED_TSMIS_TOTAL
                for section, value in arithmetic.items()
                if section != "HIGHWAY GROUP"),
            "valid section partitions did not pass")
    missing_total = dict(valid_record)
    del missing_total["Total Intersections"]
    rejected(lambda: oracle._validate_record(missing_total, "missing total"),
             "missing Total was accepted")
    all_zero = {key: 0 for key in valid_record}
    all_zero["Total Intersections"] = oracle.EXPECTED_TSMIS_TOTAL
    rejected(lambda: oracle._validate_record(all_zero, "all zero"),
             "all-zero categories with nonzero Total were accepted")
    for section, keys in oracle.SECTION_CATEGORIES.items():
        if section == "HIGHWAY GROUP":
            continue
        changed = dict(valid_record)
        changed[keys[0]] += 1
        rejected(lambda changed=changed, section=section:
                 oracle._validate_record(changed, section),
                 f"{section} partition mutation was accepted")
    highway_changed = dict(valid_record)
    highway_changed[oracle.SECTION_CATEGORIES["HIGHWAY GROUP"][0]] += 1
    oracle._validate_record(highway_changed, "documented highway exemption")

    require(oracle._resolve_rural_urban_rows([
        (None, "U-URBAN -I INSIDE CITY"),
        (5, "-O OUTSIDE CITY")]) == {"U-O": 5},
        "count-less Urban parent was not retained")
    require(oracle._resolve_rural_urban_rows([
        (1, "R-RURAL -I INSIDE CITY"), (2, "-O OUTSIDE CITY"),
        (3, "U-URBAN -I INSIDE CITY"), (4, "-O OUTSIDE CITY"),
        (5, "+-INVALID DATA")]) == {
            "R": 1, "R-O": 2, "U": 3, "U-O": 4, "+": 5},
        "valid Rural/Urban transition sequence drifted")
    rejected(lambda: oracle._resolve_rural_urban_rows([
        (5, "-O OUTSIDE CITY")]), "orphan child was accepted")
    rejected(lambda: oracle._resolve_rural_urban_rows([
        (None, "R-RURAL -I INSIDE CITY"), (1, "-O OUTSIDE CITY"),
        (2, "-O OUTSIDE CITY")]), "duplicate child was accepted")

    require(oracle._fold_tsn_control_rows([
        ("J", 1), ("K", 2), ("L", 3), ("M", 4), ("N", 5), ("P", 6)
    ]) == {"S": 21}, "distinct J-P fold did not pass")
    require(oracle._fold_tsn_control_rows([
        ("S", 7), ("J", 1), ("P", 2)]) == {"S": 10},
        "distinct source S/J/P mapping did not remain explicit")
    rejected(lambda: oracle._fold_tsn_control_rows([
        ("J", 1), ("J", 2)]), "repeated J was silently folded")
    rejected(lambda: oracle._fold_tsn_control_rows([
        ("J", 1.5)]), "fractional folded control count was accepted")

    metadata = {
        "Title": "TSMIS Reports",
        "Creator": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) HeadlessChrome/150.0.0.0 Safari/537.36"),
        "Producer": "Skia/PDF m150",
        "CreationDate": "D:20260710023441+00'00'",
        "ModDate": "D:20260710023441+00'00'",
    }
    page1 = "\n".join((
        "REPORT DATE : 07/09/2026", "REFERENCE DATE : 07/10/2026",
        "SUBMITTOR : Yunus.Shaikh@dot.ca.gov",
        "REPORT TITLE : TSAR - Intersection Summary", "ROUTE 001", "PM ALL"))
    page2 = "\n".join((
        "TSAR – Intersection Summary", "District: ALL County: ALL Route: 001",
        "Ref Date: 2026-07-10", "Total Intersections = 1",
        *(source for _section, source, _cats in oracle.SECTIONS)))
    oracle._validate_pdf_provenance(page1, page2, metadata, "001", "valid.pdf")
    rejected(lambda: oracle._validate_pdf_provenance(
        page1.replace("07/09/2026", "07/08/2026"), page2, metadata,
        "001", "date.pdf"), "PDF report-date mutation was accepted")
    rejected(lambda: oracle._validate_pdf_provenance(
        page1, page2.replace("Route: 001", "Route: 002"), metadata,
        "001", "route.pdf"), "PDF internal-route mutation was accepted")
    rejected(lambda: oracle._validate_pdf_provenance(
        page1, page2.replace("CONTROL TYPES", "CONTROL MODES"), metadata,
        "001", "header.pdf"), "PDF section-header mutation was accepted")
    bad_metadata = dict(metadata)
    bad_metadata["ModDate"] = "D:20260710023442+00'00'"
    rejected(lambda: oracle._validate_pdf_provenance(
        page1, page2, bad_metadata, "001", "metadata.pdf"),
        "PDF metadata mutation was accepted")
    words = _pdf_words()
    parsed_words = oracle._pdf_count_words(words, "valid.pdf")
    require(all(len(parsed_words[column]) == len(oracle.PDF_COLUMNS[column])
                for column in parsed_words), "valid PDF band geometry did not pass")
    moved_words = copy.deepcopy(words)
    moved_words[0]["x1"] = 257.0
    moved_words[0]["x0"] = 247.0
    rejected(lambda: oracle._pdf_count_words(moved_words, "moved.pdf"),
             "PDF band mutation was accepted")

    xlsx_record = {"001": _valid_record()}
    pdf_record = {"001": _valid_record()}
    require(oracle._cross_format(xlsx_record, pdf_record)["all_exact"],
            "valid Excel/PDF record did not pass")
    pdf_record["001"][oracle.TSMIS_ORDER[0]] += 1
    changed = oracle._cross_format(xlsx_record, pdf_record)
    require(not changed["all_exact"] and changed["difference_count"] == 1,
            "Excel/PDF single-value mutation was not isolated")

    with tempfile.TemporaryDirectory() as raw_temp:
        root = Path(raw_temp)
        valid_path = root / "valid-tsn.xlsx"
        _write_tsn(valid_path, list(oracle.TSN_ROWS))
        loaded = oracle._load_tsn_normalized(valid_path)
        require(tuple(loaded["order"]) == oracle.TSN_ORDER,
                "valid TSN normalized order did not pass")

        swapped = list(oracle.TSN_ROWS)
        swapped[0], swapped[1] = swapped[1], swapped[0]
        swapped_path = root / "swapped.xlsx"
        _write_tsn(swapped_path, swapped)
        rejected(lambda: oracle._load_tsn_normalized(swapped_path),
                 "TSN order mutation was accepted")

        duplicate = list(oracle.TSN_ROWS)
        duplicate[1] = (duplicate[0][0], duplicate[1][1])
        duplicate_path = root / "duplicate.xlsx"
        _write_tsn(duplicate_path, duplicate)
        rejected(lambda: oracle._load_tsn_normalized(duplicate_path),
                 "duplicate TSN category was accepted")

        typed = list(oracle.TSN_ROWS)
        typed[0] = (typed[0][0], True)
        typed_path = root / "typed.xlsx"
        _write_tsn(typed_path, typed)
        rejected(lambda: oracle._load_tsn_normalized(typed_path),
                 "Boolean TSN count was accepted")

        fraction = list(oracle.TSN_ROWS)
        fraction[0] = (fraction[0][0], 1.5)
        fraction_path = root / "fraction.xlsx"
        _write_tsn(fraction_path, fraction)
        rejected(lambda: oracle._load_tsn_normalized(fraction_path),
                 "fractional TSN count was accepted")

        count_drift = list(oracle.TSN_ROWS)
        count_drift[0] = (count_drift[0][0], count_drift[0][1] + 1)
        count_path = root / "count-drift.xlsx"
        _write_tsn(count_path, count_drift)
        rejected(lambda: oracle._load_tsn_normalized(count_path),
                 "TSN count mutation was accepted")

        reference = root / "reference.xlsx"
        _write_tsnr(reference)
        decision = oracle._load_tsnr_reference(reference)
        require(decision["semantic_mapping_proven"],
                "valid TSNR mapping did not pass")
        for name, old, new in (
                ("f", "F. Four-Way Flasher (Red on Mainline)",
                 "F. Four-Way Flasher (Red on All)"),
                ("g", "G. Four-Way flasher (Red on All)",
                 "G. Four-Way flasher (Red on Mainline)"),
                ("jp", "TSN: J source", "TSN: X source")):
            target = root / f"reference-{name}.xlsx"
            workbook = load_workbook(reference)
            ws = workbook["Sheet1"]
            cell = "H3" if name == "jp" else "B3"
            ws[cell] = str(ws[cell].value).replace(old, new)
            workbook.save(target)
            workbook.close()
            rejected(lambda target=target: oracle._load_tsnr_reference(target),
                     f"TSNR {name} mapping mutation was accepted")

        zip_a = root / "a.xlsx"
        zip_b = root / "b.xlsx"
        zip_c = root / "c.xlsx"
        zip_d = root / "d.xlsx"
        _write_zip(zip_a, b"core-A", b"sheet-A")
        _write_zip(zip_b, b"core-B-with-different-length", b"sheet-A")
        _write_zip(zip_c, b"core-A", b"sheet-B")
        _write_zip(zip_d, b"core-A", b"sheet-A", ("docProps/app.xml", b"app"))
        stable_a = oracle._zip_digest_without_core(zip_a)
        stable_b = oracle._zip_digest_without_core(zip_b)
        require(stable_a == stable_b
                and stable_a["excluded_members"] == ["docProps/core.xml"],
                "core.xml-only volatility changed stable package identity")
        require(stable_a["canonical_member_sha256"]
                != oracle._zip_digest_without_core(zip_c)["canonical_member_sha256"],
                "worksheet mutation escaped stable package identity")
        require(stable_a["canonical_member_sha256"]
                != oracle._zip_digest_without_core(zip_d)["canonical_member_sha256"],
                "non-core member addition escaped stable package identity")

        literal = root / "literal.xlsx"
        formula = root / "formula.xlsx"
        for path, value in ((literal, 2), (formula, "=1+1")):
            workbook = Workbook()
            workbook.active["A1"] = value
            workbook.save(path)
            workbook.close()
        require(oracle._workbook_semantic_digest(literal)[
            "ordered_semantic_sha256"] != oracle._workbook_semantic_digest(formula)[
                "ordered_semantic_sha256"],
                "formula/literal semantic mutation escaped workbook identity")

    truth = oracle._comparison_truth(
        oracle.EXPECTED_TSMIS_AGGREGATE, oracle.TSN_COUNTS)
    exact_rows = [{
        "category": row["category"], "status": row["status"],
        "tsmis": row["tsmis"], "tsn": row["tsn"],
    } for row in truth["rows"]]
    require(not oracle._semantic_gaps(exact_rows, truth["rows"]),
            "exact comparison truth was marked defective")
    wrong_value = copy.deepcopy(exact_rows)
    wrong_value[0]["tsmis"] += 1
    require(oracle._semantic_gaps(wrong_value, truth["rows"]),
            "comparison value mutation was accepted")
    wrong_status = copy.deepcopy(exact_rows)
    wrong_status[16]["status"] = "Both"
    wrong_status[16]["tsn"] = 0
    require(oracle._semantic_gaps(wrong_status, truth["rows"]),
            "structural absence was collapsed to zero")

    original_run = oracle.run
    original_current = oracle._publication_current
    original_emit = oracle._emit
    try:
        with tempfile.TemporaryDirectory() as raw_temp:
            root = Path(raw_temp)
            output = root / "decision.json"
            acceptance = output.with_suffix(output.suffix + ".acceptance.json")
            rejection = output.with_suffix(output.suffix + ".rejection.json")
            acceptance.write_text('{"accepted": true}\n', encoding="utf-8")
            oracle.run = lambda _args: copy.deepcopy(
                _publication_fixture(audit_complete=True))
            oracle._publication_current = lambda *_args: (True, {})
            oracle._emit = lambda _payload: None
            exit_code = oracle.main(["--output", str(output)])
            require(exit_code == 1,
                    "open product findings were accepted without authorization")
            require(not acceptance.exists(),
                    "rejected rerun retained a stale acceptance record")
            require(rejection.exists(), "rejected rerun wrote no rejection record")
            decision = json.loads(rejection.read_text(encoding="utf-8"))
            require(decision["accepted"] is False
                    and decision["reason"] == "open_product_findings_not_authorized",
                    "open-finding rejection decision is not explicit")

            exit_code = oracle.main([
                "--output", str(output), "--allow-open-findings"])
            require(exit_code == 0 and acceptance.exists() and not rejection.exists(),
                    "authorized open-finding decision did not accept cleanly")
            decision = json.loads(acceptance.read_text(encoding="utf-8"))
            require(decision["accepted"] is True
                    and decision["open_product_findings_authorized"] is True
                    and decision["production_comparison_semantics_exact"] is True
                    and decision["normalized_source_full_conservation"] is False,
                    "acceptance does not bind the separated outcome facts")

            oracle.run = lambda _args: copy.deepcopy(
                _publication_fixture(audit_complete=False))
            exit_code = oracle.main([
                "--output", str(output), "--allow-open-findings"])
            require(exit_code == 2 and rejection.exists() and not acceptance.exists(),
                    "audit-false publication did not fail closed")
    finally:
        oracle.run = original_run
        oracle._publication_current = original_current
        oracle._emit = original_emit

    print(
        "OK Stage-8 Intersection Summary gate: strict typed counts; exact suffix/drop/"
        "duplicate/order/170 routes; missing Total and every partition; Rural/Urban "
        "parents/orphans; distinct J-P fold and repeated-J rejection; PDF header/band/"
        "route/provenance mutations; Excel/PDF mismatch; TSN order/duplicate/type/count "
        "drift; TSNR F/G/J-P decision drift; core.xml-only package volatility; formula/"
        "literal semantics; structural absence; detached acceptance/rejection")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
