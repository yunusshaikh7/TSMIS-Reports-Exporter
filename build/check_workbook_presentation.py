"""Stored presentation and self-description of the comparison workbook.

The audit (PCOA-FINAL-008 / -009 / -014 / -019) found that a delivered workbook
can be internally correct and still unreadable or silent about itself:

  * a stored column width that cannot fit the identity in it, so two different
    rows read the same ("Ramp Type: C", "Highway Group: R");
  * a wholly-CONTEXT column reported as `0` differences, indistinguishable from
    a compared column that matched everywhere;
  * a values twin whose headline verdict is a formula with no cached value, so
    every consumer that does not recalculate reads the deliverable's single
    most important line as blank.

This check builds a summary-shaped and a detail-shaped comparison in BOTH
flavors and measures them with the AUDIT'S OWN gate — imported from the
committed `stage2-measure-clipping.py`, so the product check and the oracle
cannot drift apart — then asserts the two self-description contracts.
"""
from __future__ import annotations

import datetime
import importlib.util
from pathlib import Path

from _checklib import Checker, ROOT, scripts_path, temp_dir

scripts_path()

from compare_core import (CompareSchema, _CELL_PAD_PX,  # noqa: E402
                          _text_px, _usable_px, run_compare)
from openpyxl import load_workbook  # noqa: E402
from summary_layout import (Cat, Section, SummarySpec,  # noqa: E402
                            make_extra_sheet_writer)

GATE = (ROOT / "docs" / "planning" / "post-comparison-perfection-output-audit"
        / "stage2-measure-clipping.py")

# A statewide-summary shape: the category IS the identity, and the longest of
# these are the exact labels the audit measured at 89 px of available width.
SUMMARY_SCHEMA = CompareSchema(
    report_name="Presentation Summary", header=["Category", "Count"],
    side_a="SSOR-PROD 2026-07-23", side_b="SSOR-PROD 2026-07-09",
    id_noun="category", id_noun_plural="categories")
SUMMARY_A = [
    ["Ramp Type: C - Direct or Semi-direct Connector (Left)", "1873"],
    ["RURAL/URBAN/SUBURBAN: U-O - URBAN -O OUTSIDE CITY", "441"],
    ["Population: R-RURAL -O OUTSIDE CITY", "77"],
    ["Highway Group: R - Right", "12"],
]
SUMMARY_B = [[k, v if k.startswith("Highway Group") else str(int(v) + 1)]
             for k, v in SUMMARY_A]

# A detail shape: a composite key, a wholly-context column, a compared column
# that matches everywhere (which must still report a real 0), and a compared
# column that differs.
DETAIL_SCHEMA = CompareSchema(
    report_name="Presentation Detail",
    header=["Key", "Description", "City", "PM"],
    side_a="TSMIS (PDF)", side_b="TSMIS (Excel)",
    id_noun="location", id_noun_plural="locations",
    context_fields=("City",))
DETAIL_A = [[f"00{i} / ORA / R000.{i:03d}", f"DESCRIPTION OF LOCATION {i}",
             f"CITY {i}", f"{i}.5"] for i in range(1, 25)]
DETAIL_B = [[k, d + ("X" if i % 5 == 0 else ""), "A DIFFERENT CITY", pm]
            for i, (k, d, _c, pm) in enumerate(DETAIL_A)]

# ---- RB2-R2-001: the widest string is not the longest one -------------------
# Character count and rendered width are not order-equivalent: `WWWWWWWWWW` is
# 10 characters and 135.25 px where `iiiiiiiiiii` is 11 characters and 38.65 px.
# Any width heuristic that shortlists candidates by `len` before measuring can
# therefore discard the one value that needed the room. Every shape above has
# its widest string as its longest, which cannot catch that; each set below
# inverts the order on purpose, for the field values, the identity keys, the
# route, the one-sided keys, and the summary category labels.
WIDE_PROBE = ("WWWWWWWWWW", "iiiiiiiiiii")     # the exact strings in the return


def _wide(n):
    """`n` characters of the widest Calibri/Arial glyph."""
    return "W" * n


def _narrow(n):
    """`n` characters of the narrowest Calibri/Arial glyph."""
    return "i" * n


WIDE_SCHEMA = CompareSchema(
    report_name="Presentation Wide Glyph", header=["Key", "Value", "Tail"],
    side_a="TSMIS (PDF)", side_b="TSMIS (Excel)",
    id_noun="location", id_noun_plural="locations")
# In each of these key sets the pixel-widest member is the SHORTEST, so a
# "longest 1" or "longest 3" shortlist provably never measures it. The one-sided
# sets give each "Only in ..." sheet its own four-key column with the same trap.
WIDE_BOTH = (_wide(10), _narrow(11), _narrow(12), _narrow(13))
WIDE_ONLY_A = (_wide(9), _narrow(14), _narrow(15), _narrow(16))
WIDE_ONLY_B = (_wide(8), _narrow(17), _narrow(18), _narrow(19))
WIDE_ROUTE, NARROW_ROUTE = _wide(6), _narrow(9)


def _wide_row(key, value):
    """`[route, key, value, tail]` — the wide route rides the wide key, and the
    populated tail stops the value cell from legally spilling."""
    return [WIDE_ROUTE if key.startswith("W") else NARROW_ROUTE,
            key, value, "t"]


# Every paired row differs, so each renders an "<A> ≠ <B>" pair. Side A's widest
# value is `WWWWWWWWWW` and side B's is `MMMMMMMMMM`, while BOTH sides' longest
# values are narrow — so a character-count shortlist sizes the field for a pair
# that is far narrower than the pair actually written into it.
WIDE_A = ([_wide_row(k, k) for k in WIDE_BOTH]
          + [_wide_row(k, k) for k in WIDE_ONLY_A])
WIDE_B = ([_wide_row(k, "M" * 10 if k.startswith("W") else _narrow(len(k) + 4))
           for k in WIDE_BOTH]
          + [_wide_row(k, k) for k in WIDE_ONLY_B])

# The same inversion for summary_layout's Category column, reached through the
# production extra-sheet writer rather than by calling the renderer directly.
WIDE_CATS = (_wide(24), _narrow(40), _narrow(44), _narrow(48))
WIDE_SPEC = SummarySpec(
    report="Presentation Wide Glyph", sheet_name="Wide Categories",
    title="Wide-glyph category labels",
    sections=(Section(name=_wide(20),
                      cats=tuple(Cat(slug=f"c{i}", label=label, key=label)
                                 for i, label in enumerate(WIDE_CATS))),))
WIDE_CAT_SCHEMA = CompareSchema(
    report_name="Presentation Wide Categories", header=["Category", "Count"],
    side_a="TSMIS", side_b="TSN", id_noun="category",
    id_noun_plural="categories",
    extra_sheet_writer=make_extra_sheet_writer(WIDE_SPEC))
WIDE_CAT_A = [[label, str(10 + i)] for i, label in enumerate(WIDE_CATS)]
WIDE_CAT_B = [[label, str(11 + i)] for i, label in enumerate(WIDE_CATS)]

# ---- RB2-R2-002: a real value is wider than any product cap -----------------
# RB-2 briefly bounded identity columns at this stored width. It gave them 425
# usable pixels and clipped 4,978 actual cells across the corpus, because a cap
# is only safe where the caller WRAPS what it cannot fit and an identity column
# never wraps. The number survives HERE, and only here, so the fixture can prove
# the cap would have clipped: nothing in the product reads it any more.
RETIRED_CAP_WIDTH = 60.0

# Two real-shaped Highway Log descriptions differing by one word, so the cell
# renders the "<A> ≠ <B>" pair that the corpus showed reaching 1,069 px.
LONG_A = ("NB LNS RTE 1 OVR RTE 17 AND WBD LNS RTE 17 CONNECTOR OVERCROSSING "
          "AT PM 42.706")
LONG_B = ("NB LNS RTE 1 OVR RTE 17 AND WBD LNS RTE 17 CONNECTOR UNDERCROSSING "
          "AT PM 42.706")
# An identity that is itself past the cap, so the key column is tested too.
LONG_KEY = "LOCATION IDENTITY WIDER THAN ANY CAPPED COLUMN CAN SHOW / 000.421"
LONG_KEY_ONLY_A = "IDENTITY PRESENT ON SIDE A ONLY, ALSO PAST THE RETIRED CAP"

# Four leading fields push Description past the committed oracle's eighth
# column, which is the blind spot this fixture exists to stand in: the failing
# corpus cell was `Comparison!AI2`.
LONG_SCHEMA = CompareSchema(
    report_name="Presentation Long Value",
    header=["Key", "Near", "Mid", "Far", "Description", "Tail"],
    side_a="TSMIS", side_b="TSN",
    id_noun="location", id_noun_plural="locations")


def _long_row(key, description):
    """`[route, key, near, mid, far, description, tail]` — the populated tail
    stops the description from legally spilling into empty space."""
    return ["001", key, "n", "m", "f", description, "t"]


LONG_A_ROWS = [_long_row(LONG_KEY, LONG_A),
               _long_row(LONG_KEY_ONLY_A, LONG_A)]
LONG_B_ROWS = [_long_row(LONG_KEY, LONG_B)]

# ---- RB2-R2-002: the one bound that is NOT a product choice ------------------
# Excel rejects a stored width above this, so it is where fitting stops being
# possible rather than where the product decides to stop. A column that lands
# here has its cells WRAPPED, so all of the text still shows. No real corpus
# value reaches it — the widest measured 251.85 of 255 — which is exactly why
# the path needs a fixture: an untested branch is not a remedy.
EXCEL_MAX_COL_WIDTH = 255.0
HUGE_VALUE = (" ".join(f"SEGMENT {i:03d} OF A PATHOLOGICALLY LONG DESCRIPTION"
                       for i in range(1, 12)))
# Every writer that can put text in a ceiling-width column gets driven here, not
# just the first ones that were fixed: the Comparison and Only-in sheets, the
# per-side data sheets, Routes (a huge ROUTE id), the category sheet (a huge
# CATEGORY label plus side names long enough to burst the old fixed 13), and
# Provenance (a huge ROLE). Hand-wiring the wrap into three of these left the
# other three clipping, so the fixture covers all of them.
HUGE_ROUTE = "ROUTE " + HUGE_VALUE
HUGE_ROLE = "ROLE " + HUGE_VALUE
HUGE_CATEGORY = "CATEGORY " + HUGE_VALUE
HUGE_SIDE_A, HUGE_SIDE_B = "SSOR-PROD 2026-07-23", "SSOR-PROD 2026-07-09"
# One key on each side only, so BOTH "Only in …" sheets get a populated body
# row. Identical identities on the two sides pair perfectly and leave those
# sheets header-only, which would let their writer skip the ceiling entirely
# while the gate stayed green.
HUGE_KEY_BOTH = "KEY BOTH " + HUGE_VALUE
HUGE_KEY_ONLY_A = "KEY ONLY A " + HUGE_VALUE
HUGE_KEY_ONLY_B = "KEY ONLY B " + HUGE_VALUE

HUGE_SCHEMA = CompareSchema(
    report_name="Presentation Ceiling", header=["Key", "Description", "Tail"],
    side_a=HUGE_SIDE_A, side_b=HUGE_SIDE_B,
    id_noun="location", id_noun_plural="locations")
HUGE_A = [[HUGE_ROUTE, HUGE_KEY_BOTH, HUGE_VALUE, "t"],
          [HUGE_ROUTE, HUGE_KEY_ONLY_A, HUGE_VALUE, "t"]]
HUGE_B = [[HUGE_ROUTE, HUGE_KEY_BOTH, HUGE_VALUE + " AND THEN SOME MORE", "t"],
          [HUGE_ROUTE, HUGE_KEY_ONLY_B, HUGE_VALUE, "t"]]
HUGE_PROVENANCE = {
    "recipe": {"report": "Presentation Ceiling", "banner": "ceiling fixture"},
    "inputs": [{"role": HUGE_ROLE, "selection": "a.xlsx", "kind": "file",
                "sha256": "0" * 64}],
}

# The category sheet reads CONSOLIDATED rows as `[category, count]`, so it needs
# its own shape: fed the detail rows above it produced `[label, None, None,
# None]`, whose empty neighbour let the long label legally spill — no clipping,
# no wrap required, and the writer's ceiling path never executed.
HUGE_SPEC = SummarySpec(
    report="Presentation Ceiling", sheet_name="Ceiling Categories",
    title="Ceiling-width category labels",
    sections=(Section(name="Section",
                      cats=(Cat(slug="c0", label=HUGE_CATEGORY,
                                key=HUGE_CATEGORY),)),))
HUGE_CAT_SCHEMA = CompareSchema(
    report_name="Presentation Ceiling Categories",
    header=["Category", "Count"],
    side_a=HUGE_SIDE_A, side_b=HUGE_SIDE_B,
    id_noun="category", id_noun_plural="categories",
    extra_sheet_writer=make_extra_sheet_writer(HUGE_SPEC))
HUGE_CAT_A = [[HUGE_CATEGORY, "10"]]
HUGE_CAT_B = [[HUGE_CATEGORY, "11"]]

# ---- RB2-R2-002 round 5: a difference against an EMPTY side ----------------
# `_field_display_expr` substitutes the literal "(blank)" for an empty side, so
# this pair RENDERS as "WWWW... != (blank)" -- wider than the value alone, which
# is all the sizing used to measure when one side's widest value was empty.
BLANK_SCHEMA = CompareSchema(
    report_name="Presentation Blank Side", header=["Key", "Value", "Tail"],
    side_a="TSMIS", side_b="TSN",
    id_noun="location", id_noun_plural="locations")
BLANK_A = [["k1", "W" * 20, "t"]]
BLANK_B = [["k1", "", "t"]]

# The same defect one level subtler, and the shape the first fixture missed: a
# side that holds BOTH a short value and an empty one. Keeping a single widest
# raw value per side loses the fact that side A can also render the marker, and
# the marker is wider than "x". Row k3 adds the TRIM case -- whitespace-only is
# stored non-empty but RENDERS blank, because both flavors apply Excel TRIM.
MIXED_SCHEMA = CompareSchema(
    report_name="Presentation Mixed Blank", header=["Key", "Value", "Tail"],
    side_a="TSMIS", side_b="TSN",
    id_noun="location", id_noun_plural="locations")
MIXED_A = [["k1", "x", "t"], ["k2", "", "t"], ["k3", "   ", "t"]]
MIXED_B = [["k1", "W" * 7, "t"], ["k2", "W" * 7, "t"], ["k3", "W" * 7, "t"]]

# ---- RB2-R2-002 round 7: the DIFFERENCE font is bold ------------------------
# Conditional formatting renders every differing value bold, and bold is wider
# than regular at the same point size. Sizing measured regular, so the columns a
# comparison exists to show were the ones under-sized. The delta is ~0.7%, which
# is invisible on short text and crosses the oracle's 6px tolerance only once the
# value is long -- hence 90 characters a side, which is Codex's own repro.
BOLD_SCHEMA = CompareSchema(
    report_name="Presentation Bold Difference", header=["Key", "Value"],
    side_a="A", side_b="B", id_noun="location", id_noun_plural="locations")
BOLD_A = [["k", "A" * 90]]
BOLD_B = [["k", "B" * 90]]

# The TOTAL row writes its label bold beside a populated count, and footnote rows
# do the same -- neither was in the Category column's candidate set.
TOTAL_SPEC = SummarySpec(
    report="Presentation Total Label", sheet_name="Total Categories",
    title="Total and footnote labels are identities too",
    sections=(Section(name="S", cats=(Cat(slug="c", label="x", key="x"),)),),
    total=Cat(slug="total", label="A" * 90, key="TOTAL"))
TOTAL_SCHEMA = CompareSchema(
    report_name="Presentation Total", header=["Category", "Count"],
    side_a="TSMIS", side_b="TSN", id_noun="category",
    id_noun_plural="categories",
    extra_sheet_writer=make_extra_sheet_writer(TOTAL_SPEC))
TOTAL_A = [["x", "1"], ["TOTAL", "2"]]
TOTAL_B = [["x", "1"], ["TOTAL", "3"]]

# ---- RB2-R2-002 round 8: measured form vs SERIALIZED form -------------------
# `str(1e20)` is "1e+20"; the cell stores "100000000000000000000". Sizing that
# measures the repr under-sizes the column by 121 px of rendered text.
SERIAL_SCHEMA = CompareSchema(
    report_name="Presentation Serialized", header=["Key", "Value", "Tail"],
    side_a="A", side_b="B", id_noun="location", id_noun_plural="locations")
SERIAL_A = SERIAL_B = [["k", 1e20, "stop"]]

# ---- RB2-R2-002 round 8: a number too narrow renders ### --------------------
# 1,000 duplicate occurrences of one key, which stays under the exact-pairing
# cap. The occurrence column held a hard-coded 4 and could not show "1000".
OCC_SCHEMA = CompareSchema(
    report_name="Presentation Occurrence", header=["Key", "Value"],
    side_a="A", side_b="B", id_noun="location", id_noun_plural="locations")
OCC_A = [["duplicate", "x"] for _ in range(1000)]
OCC_B = [["duplicate", "x"]]

# ---- RB2-R2-002 round 9: characters our fonts cannot speak for -------------
# U+3000 IDEOGRAPHIC SPACE is valid input -- `_xl_trim` deliberately preserves it
# -- and Arial/Calibri render it narrow while Excel falls back to an East Asian
# font and renders it FULL WIDTH. Twenty of them under-sized a column by 62 px.
WIDECHAR_SCHEMA = CompareSchema(
    report_name="Presentation Fullwidth", header=["Key", "Value", "Tail"],
    side_a="A", side_b="B", id_noun="location", id_noun_plural="locations")
WIDECHAR_A = WIDECHAR_B = [["k", ("W\u3000" * 20) + "X", "stop"]]

# ---- RB2-R2-002 round 9: a count too big for its column --------------------
# The familiar-summary count columns were sized from the SIDE NAMES only, and
# left in General format, which degrades a large integer to "1E+09".
BIGCOUNT_SPEC = SummarySpec(
    report="Presentation Big Count", sheet_name="R9 Categories",
    title="Counts must fit and must not degrade",
    sections=(Section(name="S", cats=(Cat(slug="k", label="K", key="K"),)),))
BIGCOUNT_SCHEMA = CompareSchema(
    report_name="Presentation Big Count", header=["Category", "Count"],
    side_a="A", side_b="B", id_noun="category", id_noun_plural="categories",
    extra_sheet_writer=make_extra_sheet_writer(BIGCOUNT_SPEC))
BIGCOUNT_A = [["K", "999999999"]]
BIGCOUNT_B = [["K", "0"]]

# ---- RB2-R2-002 round 10: text our fonts cannot measure --------------------
# U+0D05 MALAYALAM LETTER A is East-Asian-Width NEUTRAL and rendered 225 px wider
# than measured; U+1F600 is WIDE and exceeded one em by 92 px. Both prove the
# same thing -- glyph coverage, not typography class, is what matters -- so both
# are fixtures and the assertion is that such text WRAPS.
UNMEASURABLE_SCHEMA = CompareSchema(
    report_name="Presentation Unmeasurable", header=["Key", "Value", "Tail"],
    side_a="A", side_b="B", id_noun="location", id_noun_plural="locations")
MALAYALAM_A = MALAYALAM_B = [["k", "\u0D05" * 20, "stop"]]
EMOJI_A = EMOJI_B = [["k", "\U0001F600" * 20, "stop"]]

# ---- RB2-R2-002 round 10: interior spaces are stored, not trimmed ----------
# `_xl_trim` collapses them; the data sheet stores the raw string. Measuring the
# trimmed form sized this column for 3 characters where the cell holds 102.
SPACES_SCHEMA = CompareSchema(
    report_name="Presentation Interior Spaces", header=["Key", "Value", "Tail"],
    side_a="A", side_b="B", id_noun="location", id_noun_plural="locations")
SPACES_A = SPACES_B = [["k", "W" + " " * 100 + "W", "stop"]]

FRESHNESS_LABEL = ("Build-time source identity and duplicate pairing snapshot "
                   "is current")
CONTEXT_TEXT = "not compared (context)"
STALE_TEXT = "REGENERATE REQUIRED"


def _load_gate():
    spec = importlib.util.spec_from_file_location("stage2_clipping_gate", GATE)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _audit_everything(gate, path):
    """`(hits, excluded, visible_sheets)` — the oracle's judgement WITHOUT its
    scan window: every visible sheet, every column, every row.

    The window is not a detail. The committed oracle scans three sheets, the
    first eight columns and the first 80 rows, and a Comparison field column
    sits far past column 8 — which is how RB2-R2-002's 4,978 clipped cells sat
    inside workbooks the oracle had just certified. Fixtures are small, so the
    permanent gate can afford to look everywhere, and looking everywhere is
    exactly what it was missing.

    Two exclusions, both because the cell is NEVER RENDERED to anyone, both
    applied AFTER the oracle judges rather than inside it, and both counted and
    named rather than quietly dropped:

      * hidden columns and rows — the versioned E/D/N/U state-mask and Med-Wid
        helper columns the workbook deliberately hides;
      * hidden and VERY-hidden sheets — `__CMP_E2_SNAPSHOT_*`, the build-time
        identity binding, which Excel will not display even on Unhide.

    Nothing else is excluded: a visible cell is in scope wherever it sits."""
    from openpyxl.utils.cell import coordinate_from_string

    rows, cols = gate.SCAN_ROWS, gate.MAX_LABEL_COL
    wb = load_workbook(path, read_only=False, data_only=True)
    hits, excluded, visible = [], [], []
    try:
        gate.SCAN_ROWS, gate.MAX_LABEL_COL = 10 ** 9, 10 ** 9
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.sheet_state != "visible":
                excluded.append({"sheet": sheet, "why": ws.sheet_state})
                continue
            visible.append(sheet)
            for hit in gate.audit_sheet(ws, path.name, sheet):
                letter, row = coordinate_from_string(hit["cell"])
                col_dim = ws.column_dimensions.get(letter)
                row_dim = ws.row_dimensions.get(row)
                if (getattr(col_dim, "hidden", False)
                        or getattr(row_dim, "hidden", False)):
                    hit["why"] = "hidden column/row"
                    excluded.append(hit)
                else:
                    hits.append(hit)
        return hits, excluded, visible
    finally:
        gate.SCAN_ROWS, gate.MAX_LABEL_COL = rows, cols
        wb.close()


def _ceiling_evidence(gate, path):
    """`{sheet -> [coordinates]}` — cells that PROVE a writer took the ceiling
    path, by being cells that would have been clipped had it not.

    A witness has to satisfy all of these, and each condition rules out a way
    the check could pass for the wrong reason:

      * in a column stored at Excel's ceiling, and that column is VISIBLE, and
        the row is visible — a hidden cell is never rendered, so wrapping it
        proves nothing;
      * a literal string, not a formula and not a number — a formula's
        displayed text is not knowable here, and a number renders rounded
        rather than clipped;
      * blocked from spilling by a populated right neighbour;
      * WRAPPED; and
      * genuinely too wide — the oracle's own measurement of the text exceeds
        the ceiling column's usable width plus its tolerance.

    That last condition is the one that matters most. Without it "Report",
    "Run" and "Note" on the Provenance sheet counted as witnesses: 51 px of
    text in a 1,790 px column, wrapped only because the whole column is, and
    incapable of clipping under any circumstances. The huge role has to be the
    witness, so the check demands a cell that could actually have failed.

    Coordinates are returned rather than a count so a failure names the cell."""
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path, read_only=False)
    found = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state != "visible":
                continue
            ceiling = {}
            for i in range(1, ws.max_column + 1):
                letter = get_column_letter(i)
                dim = ws.column_dimensions.get(letter)
                if (dim is not None and dim.width is not None
                        and float(dim.width) >= EXCEL_MAX_COL_WIDTH
                        and not getattr(dim, "hidden", False)):
                    ceiling[letter] = gate.width_to_px(float(dim.width))
            if not ceiling:
                continue
            witnesses = []
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    avail = ceiling.get(cell.column_letter)
                    if avail is None:
                        continue
                    row_dim = ws.row_dimensions.get(cell.row)
                    if getattr(row_dim, "hidden", False):
                        continue
                    value = cell.value
                    if not isinstance(value, str) or not value.strip():
                        continue          # numbers render rounded, not clipped
                    if value.startswith("="):
                        continue          # displayed text unknowable here
                    nxt = ws.cell(row=cell.row, column=cell.column + 1).value
                    if nxt is None or not str(nxt).strip():
                        continue          # free to spill; proves nothing
                    if not (cell.alignment and cell.alignment.wrap_text):
                        continue
                    font = cell.font
                    need = gate.text_px(
                        value, bool(font and font.bold),
                        float(font.size or 11) if font else 11.0) + 4
                    if need > avail + gate.TOLERANCE_PX:
                        witnesses.append(cell.coordinate)
            found[name] = witnesses
        return found
    finally:
        wb.close()


def _unwrapped_blocked_formulas(path, sheets):
    """`[(sheet, coord)]` for blocked FORMULA cells that are not wrapped.

    `_audit_everything` reads `data_only=True`, so an uncached formula reads as
    `None` and its rendered text is invisible to the oracle — which is how Spot
    Check's live cells clipped past every check here while displaying the full
    source values after recalculation. Their width cannot be measured at build
    time, so the contract is that they WRAP; this asserts it directly, on the
    formulas as written, with no recalculation needed.

    Only the grid sheets are asked. The Comparison and Only-in sheets are full
    of blocked formulas that correctly do NOT wrap: their columns were sized
    from the literal source values those formulas reproduce.

    Hidden columns and rows are excluded, matching `_audit_everything`: the
    versioned helper cells are never rendered, so requiring them to wrap would
    fail a presentation-correct build for no reader's benefit. Spill is measured
    across the RUN of empty neighbours, not just the next cell, because that is
    how Excel renders it — a cell stopped two columns over is still stopped."""
    wb = load_workbook(path, read_only=False)
    bad = []
    try:
        for name in sheets:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or not cell.value.startswith("="):
                        continue
                    col_dim = ws.column_dimensions.get(cell.column_letter)
                    row_dim = ws.row_dimensions.get(cell.row)
                    if (getattr(col_dim, "hidden", False)
                            or getattr(row_dim, "hidden", False)):
                        continue          # never rendered; nothing to clip
                    stopped = False
                    for j in range(cell.column + 1, ws.max_column + 1):
                        nxt = ws.cell(row=cell.row, column=j).value
                        if nxt is not None and str(nxt).strip():
                            stopped = True
                            break
                    if not stopped:
                        continue          # spills across the rest of the row
                    if not (cell.alignment and cell.alignment.wrap_text):
                        bad.append((name, cell.coordinate))
        return bad
    finally:
        wb.close()


def _unmeasurable_unwrapped(gate, path):
    """`[(sheet, coord, sample)]` for cells holding text our fonts cannot
    measure that were NOT wrapped.

    This replaced a check built on East Asian Width, which was the wrong
    predictor in both directions: U+0D05 is EAW=Neutral and rendered 225 px
    wider than measured, U+1F600 is EAW=Wide and exceeded the one em that model
    charged. No width constant derived from that property was going to hold,
    because the property describes typography and the problem is glyph coverage.

    So the gate asserts the CONTRACT instead of a number: text our measuring
    fonts do not cover cannot be fitted honestly, so it must wrap. That is
    checkable exactly, and it cannot be defeated by finding another character —
    which is what the previous two rounds did."""
    # The gate does its OWN coverage test. Importing `measurable_text` from the
    # product made the check circular: a mutation that disabled the product's
    # predicate disabled the gate's copy of it too, and the mutation test came
    # back with zero failures against code that had stopped wrapping entirely.
    def covered(ch, bold, size):
        font = gate.font_for(bold, size)
        return abs(font.getlength(ch) - font.getlength("￾")) > 1e-6

    wb = load_workbook(path, read_only=False, data_only=True)
    bad = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or value.isascii():
                        continue
                    font = cell.font
                    size = float(font.size or 11) if font else 11.0
                    bold = bool(font and font.bold)
                    if all(ch.isascii() or covered(ch, bold, size)
                           for ch in value):
                        continue
                    dim = ws.column_dimensions.get(cell.column_letter)
                    row_dim = ws.row_dimensions.get(cell.row)
                    if (getattr(dim, "hidden", False)
                            or getattr(row_dim, "hidden", False)):
                        continue
                    align = cell.alignment
                    if align is not None and (align.wrap_text
                                              or align.shrink_to_fit):
                        continue
                    stopped = False
                    for j in range(cell.column + 1, ws.max_column + 1):
                        nxt = ws.cell(row=cell.row, column=j).value
                        if nxt is not None and str(nxt).strip():
                            stopped = True
                            break
                    if not stopped:
                        continue      # spills across the row; shows in full
                    bad.append((name, cell.coordinate, value[:24]))
        return bad
    finally:
        wb.close()


def _hash_rendered_numbers(path):
    """`[(sheet, coord, value, short_px)]` for NUMBERS too wide for their column.

    A number that does not fit is not truncated — Excel renders `###`, showing
    the reader nothing at all. Every clipping pass in this file and in the
    committed oracle skips non-string cells, so this whole failure mode has been
    invisible to all of them (RB2-R2-002 round 8: occurrence 1,000 in a width-4
    column). A number is right-aligned and never spills, so fitting is the only
    remedy and the neighbour does not matter.

    General-format integers only. A number carrying an explicit `number_format`
    renders through that format, which this does not attempt to evaluate; those
    are reported as skipped rather than silently passed."""
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path, read_only=False, data_only=True)
    bad, skipped = [], 0
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state != "visible":
                continue
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    value = cell.value
                    if isinstance(value, bool) or not isinstance(
                            value, (int, float, datetime.date, datetime.time)):
                        continue
                    # Floats, dates and times were rejected BEFORE this check,
                    # so they were neither measured nor counted as skipped —
                    # the skip counter disclosed a hole it did not cover
                    # (RB2-R2-002 round 9).
                    # "0" renders a whole number exactly as its digits, so it is
                    # measurable here. It is listed explicitly because the
                    # summary writer now sets it on every count — without this
                    # the fix for those counts would move them OUT of this
                    # check's coverage and the gate would pass by exemption.
                    fmt = cell.number_format
                    if fmt and fmt not in ("General", "@", "0"):
                        skipped += 1
                        continue
                    letter = get_column_letter(cell.column)
                    col_dim = ws.column_dimensions.get(letter)
                    row_dim = ws.row_dimensions.get(cell.row)
                    if (getattr(col_dim, "hidden", False)
                            or getattr(row_dim, "hidden", False)):
                        continue
                    if col_dim is None or col_dim.width is None:
                        continue
                    avail = _usable_px(float(col_dim.width))
                    font = cell.font
                    if isinstance(value, (datetime.date, datetime.time)):
                        # A date is a formatted serial; without evaluating the
                        # format we cannot know its rendered text, so it is
                        # disclosed rather than measured or dropped.
                        skipped += 1
                        continue
                    shown = (str(int(value))
                             if isinstance(value, float) and value.is_integer()
                             else str(value))
                    need = _text_px(shown, bool(font and font.bold),
                                    float(font.size or 11) if font else 11.0)
                    if need > avail:
                        bad.append((name, cell.coordinate, value,
                                    round(need - avail, 1)))
        return bad, skipped
    finally:
        wb.close()


def _bold_overflow(path, sheet, header):
    """`[(coord, short_px)]` for cells that do not fit BOLD, by the PRODUCT'S own
    metric.

    The committed oracle cannot answer this. It measures Calibri only, while the
    product measures the max over its candidate fonts and so already reserves
    more than Calibri needs -- a column can overflow in the product's own model
    and still look roomy to the oracle. Measured on the round-7 repro: bold needs
    1,811.0 px of the product's metric against 1,678 px usable, while the oracle
    sees 1,362.0 px against 1,683 px and reports nothing at all.

    So this asserts the product against ITSELF: a field column must fit the text
    it stores, measured the way conditional formatting will render it."""
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb[sheet]
        col = None
        for cell in ws[1]:
            if cell.value == header:
                col = cell.column_letter
                break
        if col is None:
            return []
        dim = ws.column_dimensions.get(col)
        if dim is None or dim.width is None:
            return []
        avail = _usable_px(float(dim.width))
        bad = []
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col}{row}"]
            value = cell.value
            if not isinstance(value, str) or not value.strip():
                continue
            if cell.alignment and cell.alignment.wrap_text:
                continue      # at the ceiling and wrapped: every line shows
            need = _text_px(value, True, 10.0) + _CELL_PAD_PX
            if need > avail:
                bad.append((f"{col}{row}", round(need - avail, 1)))
        return bad
    finally:
        wb.close()


def _cf_bold_ranges(ws):
    """Ranges whose CONDITIONAL FORMATTING can render a cell bold.

    `cell.font` is the cell's base font; openpyxl does not fold a conditional
    format's differential font into it. The Comparison sheet bolds every
    differing value that way, so a checker reading only `cell.font` measures
    regular text for exactly the cells that render bold and never sees the
    shortfall (RB2-R2-002 round 7). Bold is the wider case, so treating a cell
    inside such a range as bold is the conservative reading."""
    out = []
    for fmt in ws.conditional_formatting:
        for rule in fmt.rules:
            dxf = getattr(rule, "dxf", None)
            font = getattr(dxf, "font", None) if dxf is not None else None
            if font is not None and getattr(font, "b", None):
                out.append(fmt.sqref)
                break
    return out


def _clipped_run_spill(gate, path):
    """`[(sheet, coord, short_px, text)]` — clipped cells under the REAL spill
    rule: text renders across the run of empty neighbours and stops at the first
    populated cell.

    The committed oracle asks only about `col + 1`, so a cell that reaches
    through one blank and is stopped by the next reads to it as free to spill.
    That is how `Spot Check!C10` and `F10` sat 3,449.7 px and 3,465.7 px over
    their available width while every check here passed. The oracle is a frozen
    Stage-2 artifact whose recorded numbers are cited elsewhere, so it is left
    alone and this stricter pass runs beside it; its metrics are still the
    oracle's own.

    Wrapped, hidden and never-rendered cells are excluded on the same grounds as
    `_audit_everything`."""
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path, read_only=False, data_only=True)
    hits = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state != "visible":
                continue
            bold_cf = _cf_bold_ranges(ws)
            merged = {c for rng in ws.merged_cells.ranges
                      for c in rng.cells}
            def width_px(idx):
                dim = ws.column_dimensions.get(get_column_letter(idx))
                w = dim.width if dim is not None and dim.width is not None \
                    else gate.DEFAULT_COL_WIDTH
                return gate.width_to_px(float(w))
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or not value.strip():
                        continue
                    if (cell.row, cell.column) in merged:
                        continue          # the oracle owns merged geometry
                    align = cell.alignment
                    if align is not None and (align.wrap_text
                                              or align.shrink_to_fit):
                        continue
                    col_dim = ws.column_dimensions.get(cell.column_letter)
                    row_dim = ws.row_dimensions.get(cell.row)
                    if (getattr(col_dim, "hidden", False)
                            or getattr(row_dim, "hidden", False)):
                        continue
                    avail, stopped = width_px(cell.column), False
                    for j in range(cell.column + 1, ws.max_column + 1):
                        nxt = ws.cell(row=cell.row, column=j).value
                        if nxt is not None and str(nxt).strip():
                            stopped = True
                            break
                        avail += width_px(j)
                    if not stopped:
                        continue
                    font = cell.font
                    bold = bool(font and font.bold) or any(
                        cell.coordinate in rng for rng in bold_cf)
                    need = gate.text_px(
                        value, bold,
                        float(font.size or 11) if font else 11.0) + 4
                    if need > avail + gate.TOLERANCE_PX:
                        hits.append((name, cell.coordinate,
                                     round(need - avail, 1), value[:48]))
        return hits
    finally:
        wb.close()


def _live_formula_rows_are_auto_height(path, sheets):
    """`[(sheet, row)]` for rows holding a wrapped live formula that ALSO carry
    an explicit height.

    Excel does not auto-grow a row whose height was set, so such a row shows one
    line of a wrapped formula result and hides the rest — the wrap is there and
    achieves nothing. A literal on the same row can set that height, which is
    how it happens without anyone choosing it."""
    wb = load_workbook(path, read_only=False)
    bad = []
    try:
        for name in sheets:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or not cell.value.startswith("="):
                        continue
                    if not (cell.alignment and cell.alignment.wrap_text):
                        continue
                    dim = ws.row_dimensions.get(cell.row)
                    if dim is not None and dim.height is not None:
                        bad.append((name, cell.row))
                    break
        return bad
    finally:
        wb.close()


def _column_of(path, sheet, header):
    """1-based index of `header` on `sheet`, or None."""
    wb = load_workbook(path, read_only=True)
    try:
        for row in wb[sheet].iter_rows(min_row=1, max_row=1, values_only=True):
            for i, value in enumerate(row, start=1):
                if value == header:
                    return i
        return None
    finally:
        wb.close()


def _stored_width(path, sheet, letter):
    wb = load_workbook(path, read_only=False)
    try:
        dim = wb[sheet].column_dimensions.get(letter)
        return None if dim is None else dim.width
    finally:
        wb.close()


def _summary_grid(path, data_only):
    wb = load_workbook(path, data_only=data_only, read_only=True)
    try:
        return [tuple((list(row) + [None] * 4)[:4])
                for row in wb["Summary"].iter_rows(values_only=True)]
    finally:
        wb.close()


def _per_field(rows):
    """{field label -> the DIFFERENCES BY FIELD cell} for one Summary."""
    out = {}
    seen_header = False
    for row in rows:
        if row[1] == "Field" and row[3] == "# of cells differing":
            seen_header = True
            continue
        if seen_header:
            if not row[1]:
                break
            out[row[1]] = row[3]
    return out


def main():
    c = Checker()
    gate = _load_gate()
    with temp_dir("tsmis_presentation_") as tmp:
        built = {}
        for name, sc, ra, rb in (("summary", SUMMARY_SCHEMA, SUMMARY_A, SUMMARY_B),
                                 ("detail", DETAIL_SCHEMA, DETAIL_A, DETAIL_B)):
            for mode in ("values", "formulas"):
                path = Path(tmp) / f"{name}-{mode}.xlsx"
                result = run_compare(sc, ra, rb, False, path, mode=mode,
                                     name_a="a.xlsx", name_b="b.xlsx")
                c.check(f"{name}/{mode} builds", result.status == "ok", repr(result))
                built[(name, mode)] = (path, result)

        # ---- PCOA-FINAL-008 / -009: nothing the reader needs is clipped -----
        # Measured across EVERY sheet, column and row — not the oracle's
        # three-sheet / eight-column / 80-row window, which is what let
        # RB2-R2-002 through.
        for (name, mode), (path, _result) in built.items():
            hits, skipped, sheets = _audit_everything(gate, path)
            c.check(f"{name}/{mode} has no materially clipped cell in any of "
                    f"its {len(sheets)} visible sheets ({len(skipped)} "
                    "never-rendered cells/sheets excluded)",
                    not hits,
                    "; ".join(f"{h['sheet']}!{h['cell']} short "
                              f"{h['short_by_px']}px {h['text'][:40]!r}"
                              for h in hits[:6]))

        # The identity columns are WIDENED, never wrapped away: a key must read
        # on one line in the default view.
        wb = load_workbook(built[("summary", "values")][0], read_only=False)
        try:
            comparison = wb["Comparison"]
            key_width = comparison.column_dimensions["A"].width
            key_cell = comparison["A2"]
            c.check("the category column is widened to fit its longest identity",
                    key_width is not None and key_width >= 40,
                    repr(key_width))
            c.check("an identity cell is not wrapped instead of widened",
                    not (key_cell.alignment and key_cell.alignment.wrap_text))
        finally:
            wb.close()

        # ---- RB2-R2-001: sizing measures pixels, not characters -------------
        # Guard the fixture itself first. Without a readable font file the
        # product falls back to a per-character estimate, which would make
        # width and length order-equivalent and quietly rob every assertion
        # below of its teeth.
        px_wide, px_narrow = (_text_px(t, False, 10.0) for t in WIDE_PROBE)
        c.check("the fixture's shorter string really is the pixel-wider one",
                len(WIDE_PROBE[0]) < len(WIDE_PROBE[1]) and px_wide > px_narrow,
                f"{px_wide:.2f}px for {len(WIDE_PROBE[0])} chars vs "
                f"{px_narrow:.2f}px for {len(WIDE_PROBE[1])} chars")

        wide = {}
        for mode in ("values", "formulas"):
            for label, sc, ra, rb, has_route in (
                    (f"field/{mode}", WIDE_SCHEMA, WIDE_A, WIDE_B, True),
                    (f"category/{mode}", WIDE_CAT_SCHEMA, WIDE_CAT_A,
                     WIDE_CAT_B, False)):
                path = Path(tmp) / f"wide-{label.replace('/', '-')}.xlsx"
                result = run_compare(sc, ra, rb, has_route, path, mode=mode,
                                     name_a="a.xlsx", name_b="b.xlsx")
                c.check(f"wide {label} builds", result.status == "ok",
                        repr(result))
                wide[label] = path

        # The committed oracle scans Summary / Spot Check / Comparison. The
        # one-sided and category sheets are measured with the oracle's OWN
        # sheet function, so the fix is never proved against a private ruler.
        for label, path in wide.items():
            hits, skipped, sheets = _audit_everything(gate, path)
            c.check(f"wide {label}: no cell is clipped because a shorter value "
                    f"was measured in place of the wider one (all {len(sheets)} "
                    f"visible sheets, {len(skipped)} never-rendered excluded)",
                    not hits,
                    "; ".join(f"{h['sheet']}!{h['cell']} short "
                              f"{h['short_by_px']}px {h['text'][:40]!r}"
                              for h in hits[:6]))

        # ---- RB2-R2-002: no product cap clips a value -----------------------
        # The wide-glyph fixture above proves the right candidate is MEASURED.
        # It cannot prove the measurement is honoured, because none of its
        # values is long enough to reach a cap. These are: the pair below needs
        # more pixels than the retired 60.0 cap could ever have given it, and
        # it is checked in the field, key and one-sided columns.
        long_pair = LONG_A + " ≠ " + LONG_B
        need_px = _text_px(long_pair, False, 10.0) + 4
        cap_px = round(RETIRED_CAP_WIDTH * 7) + 5
        c.check("the fixture's value really is wider than the retired cap",
                need_px > cap_px,
                f"{need_px:.0f}px needed vs {cap_px:.0f}px the {RETIRED_CAP_WIDTH} "
                "cap allowed")

        for mode in ("values", "formulas"):
            path = Path(tmp) / f"long-{mode}.xlsx"
            result = run_compare(LONG_SCHEMA, LONG_A_ROWS, LONG_B_ROWS, True,
                                 path, mode=mode, name_a="a.xlsx",
                                 name_b="b.xlsx")
            c.check(f"long/{mode} builds", result.status == "ok", repr(result))

            # The blind spot, asserted rather than described: this column is
            # past where the committed oracle stops looking, which is why the
            # unwindowed scan above is the thing that makes this check real.
            col = _column_of(path, "Comparison", "Description")
            c.check(f"long/{mode}: the failing column is past the oracle's "
                    f"{gate.MAX_LABEL_COL}-column window",
                    col is not None and col > gate.MAX_LABEL_COL,
                    f"Description at column {col}")

            hits, skipped, sheets = _audit_everything(gate, path)
            c.check(f"long/{mode}: no cell is clipped by a width cap (all "
                    f"{len(sheets)} visible sheets, {len(skipped)} "
                    "never-rendered excluded)",
                    not hits,
                    "; ".join(f"{h['sheet']}!{h['cell']} short "
                              f"{h['short_by_px']}px {h['text'][:40]!r}"
                              for h in hits[:6]))

            if col is not None:
                from openpyxl.utils import get_column_letter
                width = _stored_width(path, "Comparison", get_column_letter(col))
                c.check(f"long/{mode}: the column was widened past the retired "
                        "cap rather than clipped at it",
                        width is not None and width > RETIRED_CAP_WIDTH,
                        f"stored width {width}")

        # ---- RB2-R2-002: what happens AT Excel's ceiling --------------------
        # The remedy leaves exactly one bound in place, and a bound that clips
        # is the defect this return is about. So the ceiling wraps instead, and
        # that branch is proved here rather than asserted — no corpus value
        # reaches 255, so nothing else would ever execute it.
        huge_pair = HUGE_VALUE + " ≠ " + HUGE_VALUE + " AND THEN SOME MORE"
        ceiling_usable = round(EXCEL_MAX_COL_WIDTH * 7) + 5
        c.check("the ceiling fixture really does exceed what Excel can store",
                _text_px(huge_pair, False, 10.0) + 4 > ceiling_usable,
                f"{_text_px(huge_pair, False, 10.0) + 4:.0f}px needed vs "
                f"{ceiling_usable}px at the {EXCEL_MAX_COL_WIDTH} ceiling")

        for mode in ("values", "formulas"):
            path = Path(tmp) / f"ceiling-{mode}.xlsx"
            result = run_compare(HUGE_SCHEMA, HUGE_A, HUGE_B, True, path,
                                 mode=mode, name_a="a.xlsx", name_b="b.xlsx",
                                 provenance=HUGE_PROVENANCE)
            c.check(f"ceiling/{mode} builds", result.status == "ok", repr(result))

            cat_path = Path(tmp) / f"ceiling-cat-{mode}.xlsx"
            cat_result = run_compare(HUGE_CAT_SCHEMA, HUGE_CAT_A, HUGE_CAT_B,
                                     False, cat_path, mode=mode,
                                     name_a="a.xlsx", name_b="b.xlsx")
            c.check(f"ceiling-cat/{mode} builds", cat_result.status == "ok",
                    repr(cat_result))

            # Each writer must be SHOWN to have written a wrapped ceiling cell.
            # A sheet-name check would pass on an empty Only-in sheet and on a
            # category row whose neighbour is blank — both of which leave the
            # ceiling path unexecuted.
            evidence = _ceiling_evidence(gate, path)
            evidence.update(_ceiling_evidence(gate, cat_path))
            only_in = sorted(s for s in evidence if s.startswith("Only in"))
            # The per-side DATA sheets belong here too: the category fixture
            # reaches the ceiling on their B column, so they are a writer that
            # can clip, not just a copy of the inputs.
            want = (["Comparison", "Routes", "Provenance",
                     HUGE_SPEC.sheet_name, HUGE_SIDE_A, HUGE_SIDE_B] + only_in)
            c.check(f"ceiling/{mode}: both one-sided sheets carry a body row",
                    len(only_in) == 2, f"one-sided sheets seen: {only_in}")
            missing = [s for s in want if not evidence.get(s)]
            c.check(f"ceiling/{mode}: every writer wrote a wrapped ceiling cell "
                    f"that would otherwise have clipped ({len(want)} sheets)",
                    not missing,
                    f"no qualifying witness on {missing}; witnesses "
                    + "; ".join(f"{s}:{v[:2]}" for s, v in sorted(evidence.items())))

            # Spot Check renders its values through FORMULAS, so no build-time
            # width can fit them and `data_only=True` hides them from the
            # oracle entirely. Their contract is to wrap; asserted here on the
            # formulas as written (RB2-R2-002 round 4).
            for label, target in ((f"ceiling/{mode}", path),
                                  (f"ceiling-cat/{mode}", cat_path)):
                live = _unwrapped_blocked_formulas(target,
                                                   ("Summary", "Spot Check"))
                c.check(f"{label}: every live grid formula that cannot spill "
                        "is wrapped, since its result cannot be measured",
                        not live,
                        "; ".join(f"{s}!{coord}" for s, coord in live[:8]))

            # Excel renders across a RUN of blanks; the committed oracle asks
            # only about the next cell, so this stricter pass runs beside it.
            for label, target in ((f"ceiling/{mode}", path),
                                  (f"ceiling-cat/{mode}", cat_path)):
                run_hits = _clipped_run_spill(gate, target)
                c.check(f"{label}: nothing is clipped once spill is measured "
                        "across the whole run of blanks, not just one cell",
                        not run_hits,
                        "; ".join(f"{s}!{coord} short {short}px {text!r}"
                                  for s, coord, short, text in run_hits[:6]))
                tall = _live_formula_rows_are_auto_height(
                    target, ("Summary", "Spot Check"))
                c.check(f"{label}: every row with a wrapped live formula keeps "
                        "Excel's automatic height, so the wrap can grow",
                        not tall,
                        "; ".join(f"{s} row {r}" for s, r in tall[:8]))

            col = _column_of(path, "Comparison", "Description")
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(col) if col else None
            width = _stored_width(path, "Comparison", letter) if letter else None
            c.check(f"ceiling/{mode}: the column stops at Excel's own limit, "
                    "not below it",
                    width is not None and abs(width - EXCEL_MAX_COL_WIDTH) < 1e-6,
                    f"stored width {width} in column {letter}")

            wb = load_workbook(path, read_only=False)
            try:
                cell = wb["Comparison"][f"{letter}2"]
                c.check(f"ceiling/{mode}: the cell WRAPS rather than being cut "
                        "off, so every line still shows",
                        bool(cell.alignment and cell.alignment.wrap_text),
                        repr(cell.alignment))
            finally:
                wb.close()

            for label, target in ((f"ceiling/{mode}", path),
                                  (f"ceiling-cat/{mode}", cat_path)):
                hits, skipped, sheets = _audit_everything(gate, target)
                c.check(f"{label}: nothing is reported clipped (all "
                        f"{len(sheets)} visible sheets, {len(skipped)} "
                        "never-rendered excluded)",
                        not hits,
                        "; ".join(f"{h['sheet']}!{h['cell']} short "
                                  f"{h['short_by_px']}px {h['text'][:40]!r}"
                                  for h in hits[:6]))

        # ---- RB2-R2-002 round 5: sizing must measure the (blank) marker -----
        blank_widths = {}
        shapes = (("blank", BLANK_SCHEMA, BLANK_A, BLANK_B),
                  ("mixed", MIXED_SCHEMA, MIXED_A, MIXED_B),
                  ("bold", BOLD_SCHEMA, BOLD_A, BOLD_B),
                  ("total", TOTAL_SCHEMA, TOTAL_A, TOTAL_B))
        for mode in ("values", "formulas"):
          for tag, schema, ra, rb in shapes:
            path = Path(tmp) / f"{tag}-{mode}.xlsx"
            result = run_compare(schema, ra, rb, False, path,
                                 mode=mode, name_a="a.xlsx", name_b="b.xlsx")
            c.check(f"{tag}/{mode} builds", result.status == "ok", repr(result))
            hits, _skipped, _sheets = _audit_everything(gate, path)
            run_hits = _clipped_run_spill(gate, path)
            c.check(f"{tag}/{mode}: the field is sized for the text the cell "
                    "actually renders, marker included",
                    not hits and not run_hits,
                    "; ".join(f"{h['sheet']}!{h['cell']} short "
                              f"{h['short_by_px']}px" for h in hits[:4])
                    + "; ".join(f"{s}!{coord} short {short}px"
                                for s, coord, short, _x in run_hits[:4]))
            # Conditional formatting renders a differing value BOLD, so
            # the column must fit it bold -- checked against the
            # product's own metric, the only one that can see it.
            if mode == "values":
                over = _bold_overflow(path, "Comparison",
                                      schema.header[-1])
                c.check(f"{tag}: the field column fits its value rendered "
                        "BOLD, the way a difference is displayed",
                        not over,
                        "; ".join(f"Comparison!{coord} short {short}px"
                                  for coord, short in over[:4]))
            vcol = (_column_of(path, "Comparison", "Value")
                    or _column_of(path, "Comparison", "Category"))
            from openpyxl.utils import get_column_letter as _gcl
            blank_widths[(tag, mode)] = _stored_width(
                path, "Comparison", _gcl(vcol))

        # The formulas twin writes this cell as a FORMULA, which `data_only`
        # reads as None — so the scan above can only see the values twin. The
        # two flavors are built to identical physical geometry, so asserting the
        # stored widths are equal carries the values-twin proof across to it.
        for tag, _s, _a, _b in shapes:
            c.check(f"{tag}: both flavors store the same field width, so "
                    "the values-twin measurement covers the formulas twin",
                    blank_widths.get((tag, "values"))
                    == blank_widths.get((tag, "formulas")),
                    repr({k: v for k, v in blank_widths.items()
                          if k[0] == tag}))

        # ---- RB2-R2-002 round 8: serialized form, and numbers as ### --------
        for tag, schema, ra, rb in (("serial", SERIAL_SCHEMA, SERIAL_A, SERIAL_B),
                                    ("occ", OCC_SCHEMA, OCC_A, OCC_B),
                                    ("widechar", WIDECHAR_SCHEMA,
                                     WIDECHAR_A, WIDECHAR_B),
                                    ("bigcount", BIGCOUNT_SCHEMA,
                                     BIGCOUNT_A, BIGCOUNT_B),
                                    ("malayalam", UNMEASURABLE_SCHEMA,
                                     MALAYALAM_A, MALAYALAM_B),
                                    ("emoji", UNMEASURABLE_SCHEMA,
                                     EMOJI_A, EMOJI_B),
                                    ("spaces", SPACES_SCHEMA,
                                     SPACES_A, SPACES_B)):
            for mode in ("values", "formulas"):
                path = Path(tmp) / f"{tag}-{mode}.xlsx"
                result = run_compare(schema, ra, rb, False, path, mode=mode,
                                     name_a="a.xlsx", name_b="b.xlsx")
                c.check(f"{tag}/{mode} builds", result.status == "ok",
                        repr(result))
                hits, _sk, _sh = _audit_everything(gate, path)
                run_hits = _clipped_run_spill(gate, path)
                c.check(f"{tag}/{mode}: no text cell is clipped",
                        not hits and not run_hits,
                        "; ".join(f"{h['sheet']}!{h['cell']} short "
                                  f"{h['short_by_px']}px" for h in hits[:3])
                        + "; ".join(f"{s}!{coord} short {short}px"
                                    for s, coord, short, _x in run_hits[:3]))
                wide = _unmeasurable_unwrapped(gate, path)
                c.check(f"{tag}/{mode}: text our fonts cannot measure is "
                        "WRAPPED rather than given a guessed width",
                        not wide,
                        "; ".join(f"{s}!{coord} {sample!r}"
                                  for s, coord, sample in wide[:4]))
                nums, skipped = _hash_rendered_numbers(path)
                c.check(f"{tag}/{mode}: no NUMBER is too wide for its column, "
                        f"which would render ### ({skipped} formatted skipped)",
                        not nums,
                        "; ".join(f"{s}!{coord}={val} short {short}px"
                                  for s, coord, val, short in nums[:4]))

        # ---- RB2-R2-002 round 9: the Source Files companion sheet -----------
        # Built through the PRODUCTION writer, which declared no widths at all,
        # so its own header "Route (as compared)" clipped on every workbook.
        from compare_tsn_common import write_source_files_sheet
        from openpyxl import Workbook

        sf = Path(tmp) / "source-files.xlsx"
        wb = Workbook(write_only=True)
        write_source_files_sheet(wb, [
            ("SSOR-PROD 2026-07-23", [["001", "k", "v"]],
             ["highway_log_route_001.xlsx"])])
        wb.save(sf)
        hits, _sk, _sh = _audit_everything(gate, sf)
        run_hits = _clipped_run_spill(gate, sf)
        c.check("source-files: the companion sheet fits its own headers "
                "and values",
                not hits and not run_hits,
                "; ".join(f"{h['sheet']}!{h['cell']} short "
                          f"{h['short_by_px']}px {h['text'][:32]!r}"
                          for h in hits[:4])
                + "; ".join(f"{s}!{coord} short {short}px"
                            for s, coord, short, _x in run_hits[:4]))

        # ---- PCOA-FINAL-014: a wholly-context column says so ----------------
        for mode in ("values", "formulas"):
            fields = _per_field(_summary_grid(built[("detail", mode)][0], False))
            c.check(f"detail/{mode} renders a wholly-context column as context",
                    fields.get("City") == CONTEXT_TEXT, repr(fields))
            c.check(f"detail/{mode} still counts a compared column that differs",
                    fields.get("Description") not in (None, CONTEXT_TEXT, ""),
                    repr(fields))
            pm = fields.get("PM")
            c.check(f"detail/{mode} still reports a real 0 for a compared "
                    "column with no differences",
                    (pm == 0) if mode == "values"
                    else (isinstance(pm, str) and pm.startswith("=")),
                    repr(fields))

        # ---- PCOA-FINAL-019: the values headline is readable, and stale
        #      inputs still decertify the workbook ---------------------------
        for name in ("summary", "detail"):
            path, result = built[(name, "values")]
            cached = _summary_grid(path, True)[2][1]
            typed = result.comparison_outcome
            c.check(f"{name}: the values headline is non-empty read data_only",
                    isinstance(cached, str) and cached.strip(), repr(cached))
            c.check(f"{name}: the values headline matches the typed outcome",
                    isinstance(cached, str)
                    and (cached.startswith("✓") if typed.verdict == "match"
                         else cached.startswith("✗"))
                    and (typed.verdict == "match"
                         or f"{typed.counts.differing_cells:,}" in cached),
                    f"{cached!r} vs {typed.verdict!r} / {typed.counts!r}")

            stored = _summary_grid(path, False)
            fresh = [row[2] for row in stored if row[1] == FRESHNESS_LABEL]
            c.check(f"{name}: the values twin still fails closed when stale",
                    len(fresh) == 1 and isinstance(fresh[0], str)
                    and fresh[0].startswith("=IF(")
                    and STALE_TEXT in fresh[0]
                    and "__CMP_E2_SNAPSHOT_A" in fresh[0]
                    and "__CMP_E2_SNAPSHOT_B" in fresh[0],
                    repr(fresh))
            c.check(f"{name}: the values twin discloses its stored headline",
                    any(isinstance(row[1], str)
                        and "STORED build-time result" in row[1]
                        for row in stored),
                    "no note names the stored verdict")

            formulas_path = built[(name, "formulas")][0]
            live = _summary_grid(formulas_path, False)[2][1]
            live_fresh = [row[2] for row in _summary_grid(formulas_path, False)
                          if row[1] == FRESHNESS_LABEL]
            c.check(f"{name}: the formulas twin keeps its live guarded verdict",
                    isinstance(live, str) and live.startswith("=IF(")
                    and STALE_TEXT in live, repr(live))
            c.check(f"{name}: the formulas twin's freshness row is live too",
                    len(live_fresh) == 1 and STALE_TEXT in str(live_fresh[0]),
                    repr(live_fresh))
    return c.summary()


if __name__ == "__main__":
    raise SystemExit(main())
