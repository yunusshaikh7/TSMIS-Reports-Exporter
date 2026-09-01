"""Golden check for TSN normalized-workbook identity DETERMINISM (PCOA-FINAL-017).

Force-rebuilding the TSN library from raw whose `tsn_raw_manifest.sha256` and
`normalization_version` had NOT changed produced a different
`tsn_normalized_workbook_identity` and `tsn_artifact_identity_token` for every
dataset. That token binds a committed comparison generation to its TSN source, so
pressing Settings' *Rebuild* silently invalidated every existing vs-TSN
comparison even when nothing had changed, forcing a full statewide re-comparison.

ESTABLISHED root cause (the finding labelled the openpyxl explanation an explicit
hypothesis; this check is the measurement, not the guess). openpyxl stamps
`docProps/core.xml` with `dcterms:created` / `dcterms:modified` = the wall clock
at save. Two saves of the same workbook object one second apart differ in exactly
those two elements and in nothing else — identical member names, identical ZIP
`date_time` fields, identical every other part — and two saves inside the same
second were already byte-identical. So the bytes were a function of the CLOCK,
and the identity is a sha256 over the bytes. No parser and no projection is
involved, which is why the fix moves no normalized content.

Locks:
  * two consecutive `build_consolidated(report, force=True)` calls over unchanged
    raw, straddling a clock tick, produce a byte-identical workbook and identical
    `tsn_normalized_workbook_identity` / `tsn_artifact_identity_token` — driven
    through the SHIPPED library entry point, with the real writer, marker sheet,
    atomic save, raw certificate and identity computation;
  * a rebuild that SHOULD change identity still does — changed raw bytes and a
    bumped `normalization_version` both move the token. Determinism must not
    become blindness;
  * the two district-PDF writers, which do not go through `build_normalized`,
    are byte-stable at their own save boundaries;
  * EVERY registered TSN dataset reaches a stamped writer — asserted from
    `report_catalog.TSN`, so a twelfth dataset on a new save path fails here.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_tsn_identity_determinism.py
"""
import importlib
import inspect
import io
import os
import shutil
import sys
import tempfile
import time
import zipfile
from contextlib import contextmanager
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import artifact_store
import paths
import report_catalog
import tsn_library
from events import Events

_fail = []

# openpyxl's document timestamp has one-second resolution, so two saves inside
# one second were byte-identical even PRE-fix. Every determinism assertion here
# must straddle a tick or it would pass by luck.
TICK_S = 1.1

# The three SAVE boundaries every registered dataset reaches, and how each one
# asks for stable bytes. Asserted against report_catalog.TSN below.
_STABLE_SAVES = {
    "tsn_library": ("build_normalized", "atomic_save_if"),
    "consolidate_tsn_highway_sequence": ("_write_workbook", "atomic_save_if"),
    "consolidate_tsn_highway_log": ("consolidate", "consolidate_xlsx"),
}
# The two Clean Road slots have NO normalizer by design (DEF-05) — their builders
# refuse and write nothing, so they have no bytes to make deterministic.
_NO_NORMALIZER = ("clean_intersection", "clean_ramp")


def check(name, cond, detail=""):
    suffix = f"  -> {detail}" if (not cond and detail) else ""
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}{suffix}")
    if not cond:
        _fail.append(name)


@contextmanager
def _patch(obj, name, value):
    original = getattr(obj, name)
    setattr(obj, name, value)
    try:
        yield
    finally:
        setattr(obj, name, original)


def _differing_members(a, b):
    with zipfile.ZipFile(io.BytesIO(a)) as za, zipfile.ZipFile(io.BytesIO(b)) as zb:
        names = sorted(set(za.namelist()) | set(zb.namelist()))
        return [n for n in names
                if (n not in za.namelist() or n not in zb.namelist()
                    or za.read(n) != zb.read(n))]


def _rows_sheet(data, sheet_index=0):
    """Every cell of one sheet, so 'the content did not move' is asserted on the
    values rather than inferred from the digest."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[sheet_index]]
        return [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# 1. The shipped path: build_consolidated twice over unchanged raw
# --------------------------------------------------------------------------- #
_FIXTURE_ROWS = [
    ["001", "R001.000", "A RAMP", 1, "01", "DN", "R"],
    ["002", "R002.500", "B RAMP", 2, "02", "SIS", ""],
]


def _ramp_detail_projection(_raw_path):
    """A constant projection in the shape `tsn_load_ramp_detail._project` builds
    from the statewide workbook: (rows, per-row district/county/suffix). The raw
    file therefore never has to be a real 15k-row TSN dump, while every byte the
    library writes still comes from the real writer."""
    rows = [list(r[:4]) for r in _FIXTURE_ROWS]
    dcr = [(r[4], r[5], r[6]) for r in _FIXTURE_ROWS]
    return rows, dcr


@contextmanager
def _sandboxed_library(root):
    """Hermetic: the library root is redirected, so no real staged library is
    read or written and the check leaves nothing behind."""
    saved = paths.TSN_LIBRARY_ROOT
    paths.TSN_LIBRARY_ROOT = Path(root)
    try:
        yield
    finally:
        paths.TSN_LIBRARY_ROOT = saved


def _build_once(report):
    res = tsn_library.build_consolidated(report, events=Events(), force=True)
    out = tsn_library.consolidated_path(report)
    st = tsn_library.status(report)
    return res, (out.read_bytes() if out.exists() else b""), st


def test_double_rebuild_is_deterministic(tmp):
    print("two force rebuilds over unchanged raw are byte-identical (shipped path):")
    import tsn_load_ramp_detail as rd_load
    report = "ramp_detail"
    with _sandboxed_library(tmp / "lib"):
        raw = tmp / "TSN ramp detail.xlsx"
        _write_stub_xlsx(raw, "raw generation 1")
        tsn_library.import_raw(report, [str(raw)])
        with _patch(rd_load, "tsn_rows_with_dcr", _ramp_detail_projection):
            first_res, first_bytes, first_status = _build_once(report)
            time.sleep(TICK_S)
            second_res, second_bytes, second_status = _build_once(report)

            check("both builds returned ok",
                  first_res.status == "ok" and second_res.status == "ok",
                  f"{first_res.status!r} / {second_res.status!r} "
                  f"{second_res.message!r}")
            check("the two normalized workbooks are byte-identical",
                  first_bytes == second_bytes and first_bytes,
                  f"differing members: "
                  f"{_differing_members(first_bytes, second_bytes) if first_bytes else 'no output'}")
            check("tsn_normalized_workbook_identity is unchanged",
                  first_status["normalized_workbook_identity"] == second_status["normalized_workbook_identity"],
                  f"{first_status['normalized_workbook_identity']} != "
                  f"{second_status['normalized_workbook_identity']}")
            check("tsn_artifact_identity_token is unchanged",
                  first_status["identity_token"] == second_status["identity_token"]
                  and first_status["identity_token"],
                  f"{first_status['identity_token']} != {second_status['identity_token']}")
            check("the rebuilt library is still current/reusable",
                  second_status["current"] is True, repr(second_status))

            # 3. Determinism must not become blindness.
            print("...and a rebuild that SHOULD change identity still does:")
            time.sleep(TICK_S)
            changed_rows = [list(r) for r in _FIXTURE_ROWS]
            changed_rows[0][2] = "A RAMP (RENAMED)"
            with _patch(rd_load, "tsn_rows_with_dcr",
                        lambda _p: ([r[:4] for r in changed_rows],
                                    [(r[4], r[5], r[6]) for r in changed_rows])):
                _res, content_bytes, content_status = _build_once(report)
            check("changed normalized CONTENT -> different workbook identity",
                  content_status["normalized_workbook_identity"] != first_status["normalized_workbook_identity"])
            check("changed normalized CONTENT -> different identity token",
                  content_status["identity_token"] != first_status["identity_token"])

    # A changed RAW manifest must move the token even when the projection is
    # constant — the certificate, not the bytes, carries the source claim.
    with _sandboxed_library(tmp / "lib2"):
        raw2 = tmp / "TSN ramp detail.xlsx"
        _write_stub_xlsx(raw2, "raw generation 2 — different bytes")
        tsn_library.import_raw(report, [str(raw2)])
        with _patch(rd_load, "tsn_rows_with_dcr", _ramp_detail_projection):
            _res, raw_bytes_out, raw_status = _build_once(report)
    check("changed RAW bytes -> different identity token (same projection)",
          raw_status["identity_token"] != first_status["identity_token"],
          "a new raw source must not reuse the old binding")
    check("...while the normalized CONTENT is unchanged by the raw swap",
          _rows_sheet(raw_bytes_out) == _rows_sheet(first_bytes))


def _write_stub_xlsx(path, marker):
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.append(["raw", marker])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# --------------------------------------------------------------------------- #
# 2. The two district-PDF writers, at their own save boundaries
# --------------------------------------------------------------------------- #
def test_district_writers_are_byte_stable(tmp):
    print("the two district-PDF save paths are byte-stable across a clock tick:")
    import consolidate_tsn_highway_log as hl_tsn
    import consolidate_tsn_highway_sequence as hsl_tsn
    from openpyxl import Workbook

    def _row(county, pm, description):
        return {"county": county, "pm": pm, "city": "", "hg": "D", "ft": "H",
                "dist": "0.500", "description": description}

    rows = [("001", _row("DN", "001.000", "A JCT")),
            ("002", _row("SIS", "002.000", "B JCT"))]
    out_a, out_b = tmp / "hsl_a.xlsx", tmp / "hsl_b.xlsx"
    hsl_tsn._write_workbook(rows, out_a)
    time.sleep(TICK_S)
    hsl_tsn._write_workbook(rows, out_b)
    check("consolidate_tsn_highway_sequence._write_workbook: identical bytes",
          out_a.read_bytes() == out_b.read_bytes(),
          _zip_delta(out_a.read_bytes(), out_b.read_bytes()))

    # The TSN Highway Log build saves through consolidate_xlsx; exercise the same
    # decorated workbook over the same stable save boundary.
    def decorated(path):
        wb = Workbook()
        wb.active.append(["Route", "Location"])
        wb.active.append(["001", "R001.000"])
        hl_tsn._decorate_normalized(wb)
        artifact_store.atomic_save_if(wb, path, lambda: True, stable_identity=True)
        return path.read_bytes()

    one = decorated(tmp / "hl_a.xlsx")
    time.sleep(TICK_S)
    two = decorated(tmp / "hl_b.xlsx")
    check("consolidate_tsn_highway_log's decorated workbook: identical bytes",
          one == two, _zip_delta(one, two))

    # And the DEFAULT save must be untouched: every other consolidator still
    # writes exactly as it did, so this fix moves no other family's bytes.
    def default_saved(path):
        wb = Workbook()
        wb.active.append(["Route", "Location"])
        artifact_store.atomic_save_if(wb, path, lambda: True)
        return path.read_bytes()

    plain_a = default_saved(tmp / "plain_a.xlsx")
    time.sleep(2.2)                       # ZIP timestamps are 2-second resolution
    plain_b = default_saved(tmp / "plain_b.xlsx")
    check("the DEFAULT save is still openpyxl's own (time-varying, unchanged)",
          plain_a != plain_b,
          "stable identity leaked into producers that did not opt in")


def _zip_delta(a, b):
    """Why two archives differ: content, member set, or entry timestamps."""
    if a == b:
        return ""
    with zipfile.ZipFile(io.BytesIO(a)) as za, zipfile.ZipFile(io.BytesIO(b)) as zb:
        if za.namelist() != zb.namelist():
            return f"member sets differ: {za.namelist()} vs {zb.namelist()}"
        content = [n for n in za.namelist() if za.read(n) != zb.read(n)]
        stamps = [n for n in za.namelist()
                  if za.getinfo(n).date_time != zb.getinfo(n).date_time]
        return f"differing content: {content}; differing entry timestamps: {stamps}"


# --------------------------------------------------------------------------- #
# 3. Coverage: every registered dataset reaches a stamped writer
# --------------------------------------------------------------------------- #
def test_every_dataset_reaches_a_stamped_writer():
    print("every registered TSN dataset saves through a stable-identity boundary:")
    for module, (func_name, via) in _STABLE_SAVES.items():
        mod = importlib.import_module(module)
        fn = getattr(mod, func_name, None)
        check(f"{module}.{func_name} exists", callable(fn))
        if not callable(fn):
            continue
        src = inspect.getsource(fn)
        check(f"{module}.{func_name} asks {via} for stable_identity",
              "stable_identity=True" in src and via in src,
              "its saved bytes would still depend on the wall clock")
    for spec in report_catalog.tsn_entries():
        mod_name, func_name = spec.builder.split(":")
        mod = importlib.import_module(mod_name)
        builder = getattr(mod, func_name, None)
        check(f"{spec.subdir}: builder {spec.builder} is importable", callable(builder))
        if not callable(builder):
            continue
        if spec.subdir in _NO_NORMALIZER:
            # DEF-05: reserved slots with no normalizer. They must REFUSE rather
            # than write an undetermined workbook — asserted, not assumed.
            res = builder(Path("."), Path("."))
            check(f"{spec.subdir}: has no normalizer and refuses (DEF-05)",
                  getattr(res, "status", None) == "error",
                  f"it returned {getattr(res, 'status', res)!r}; if it now writes "
                  "a workbook it needs a stable-identity save boundary")
            continue
        # A single-file normalizer delegates its writing to build_normalized; the
        # two district-PDF builders own their save. Either way the owning module
        # must be one this check has just proven asks for a stable save.
        owner = mod_name if mod_name in _STABLE_SAVES else "tsn_library"
        delegates = (owner != "tsn_library"
                     or "build_normalized" in inspect.getsource(builder))
        check(f"{spec.subdir}: reaches a stable-identity save ({owner})",
              owner in _STABLE_SAVES and delegates,
              f"{spec.builder} neither owns a stable save boundary nor delegates "
              "to tsn_library.build_normalized")


def test_the_stamp_is_a_declared_constant():
    print("the stamp is one declared constant, not 'whatever the first save saw':")
    from openpyxl import Workbook
    wb = artifact_store.stable_document_identity(Workbook())
    check("created == modified == the declared constant",
          wb.properties.created == wb.properties.modified
          and wb.properties.created.isoformat()
          == artifact_store.STABLE_DOCUMENT_TIMESTAMP,
          f"created={wb.properties.created!r} modified={wb.properties.modified!r}")
    check("the machine name is not written into the document",
          not wb.properties.lastModifiedBy,
          f"lastModifiedBy={wb.properties.lastModifiedBy!r}")
    buf = io.BytesIO()
    wb.save(buf)
    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
        core = z.read("docProps/core.xml").decode("utf-8", "replace")
    check("docProps/core.xml carries no current-year wall clock",
          str(time.gmtime().tm_year) not in core, f"core.xml={core!r}")


def main():
    print("=== TSN normalized-workbook identity determinism (PCOA-FINAL-017) ===")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_tsn_determinism_"))
    try:
        test_double_rebuild_is_deterministic(tmp)
        test_district_writers_are_byte_stable(tmp)
        test_every_dataset_reaches_a_stamped_writer()
        test_the_stamp_is_a_declared_constant()
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL TSN IDENTITY-DETERMINISM CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
