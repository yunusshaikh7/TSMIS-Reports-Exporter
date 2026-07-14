#!/usr/bin/env python3
"""Permanent adversarial gate for the independent Stage-8 Ramp Detail oracle."""

from __future__ import annotations

from datetime import date
from pathlib import Path
import sys
import tempfile
import zipfile

from openpyxl import Workbook, load_workbook


BUILD_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(BUILD_ROOT))
import phase8_ramp_detail_comparison as oracle  # noqa: E402


passed = 0


def require(condition: bool, message: str) -> None:
    global passed
    if not condition:
        raise AssertionError(message)
    passed += 1


def rejects(action, message: str) -> None:
    try:
        action()
    except oracle.AuditError:
        require(True, message)
    else:
        raise AssertionError(message)


def row(reference: str, *, source: str = "synthetic", route: str = "001",
        county: str = "LA", district: str = "07", pm: str = "1.000",
        pr: str = "", description: str = "X", hg: str = "D",
        area4: str = "N", city: str = "LA", ru: str = "U",
        raw_description: str | None = None, raw_area4: str | None = None,
        pm_suffix: str = "") -> oracle.RampRow:
    return oracle.RampRow(
        source=source, member=f"{reference}.xlsx", source_record="row 2",
        route=route, district=district, county=county, pr=pr, pm=pm,
        record_date="2020-01-01", hg=hg, area4=area4, city=city, ru=ru,
        description=description, pm_suffix=pm_suffix,
        raw_location=f"{district}-{county}-{route}",
        raw_area4=area4 if raw_area4 is None else raw_area4,
        raw_description=(description if raw_description is None
                         else raw_description))


def write_normalized(path: Path, *, mutate=None) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = oracle.TSN_NORMALIZED_SHEET
    worksheet.append(list(oracle.TSN_NORMALIZED_HEADER))
    worksheet.append([
        "001", "R", "1.2", date(2020, 1, 2), "D", "N", "LA", "U",
        "DESC", "RAMP", "O", "C", "100", "07", "LA",
    ])
    if mutate is not None:
        mutate(workbook, worksheet)
    workbook.save(path)
    workbook.close()


def write_tsmis(path: Path, *, mutate=None) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = oracle.TSMIS_SHEET
    worksheet.append(list(oracle.TSMIS_HEADER))
    worksheet.append([
        "07-LA-001", "R", "1.2", date(2020, 1, 2), None, "D", "N",
        "LA", "U", "001/DESC", None,
    ])
    if mutate is not None:
        mutate(workbook, worksheet)
    workbook.save(path)
    workbook.close()


def word(text: str, x0: float, *, top: float = 10.0,
         width: float = 5.0) -> dict[str, object]:
    return {
        "text": text, "x0": x0, "x1": x0 + width,
        "top": top, "bottom": top + 5.0,
    }


def valid_header_words(top: float = 10.0) -> list[dict[str, object]]:
    return [
        word("LOCATION", 0, top=top, width=30), word("P", 35, top=top),
        word("PM", 50, top=top, width=12), word("RECORD", 100, top=top, width=30),
        word("AREA", 160, top=top, width=20), word("CITY", 210, top=top, width=18),
        word("CODE", 232, top=top, width=20), word("R", 270, top=top),
        word("O", 290, top=top), word("T", 310, top=top),
        word("DESCRIPTION", 350, top=top, width=50),
    ]


def main() -> int:
    require(oracle._pm("009.600") == "9.600", "valid PM normalization drift")
    require(oracle._pm("0") == "0.000", "zero PM normalization drift")
    rejects(lambda: oracle._pm("+1.000"), "leading-plus PM accepted")
    rejects(lambda: oracle._pm("1.0004"), "over-precision PM accepted")
    rejects(lambda: oracle._pm("-1"), "negative PM accepted")
    rejects(lambda: oracle._date("2026-02-30"), "impossible date accepted")
    require(oracle._route("5s") == "005S", "suffix route normalization drift")

    require(oracle._tsmis_description("001/DESC", "001", pdf=False)[0] == "DESC",
            "exact outer route was not stripped")
    require(oracle._tsmis_description("005/DESC", "005S", pdf=False)[0] == "DESC",
            "base outer route on suffix route was not stripped")
    require(oracle._tsmis_description("505/128/RUSSELL", "505", pdf=False)[0]
            == "128/RUSSELL", "different inner numeric source prefix was lost")
    require(oracle._tsmis_description("128/RUSSELL", "505", pdf=False)[0]
            == "128/RUSSELL", "different numeric prefix was mistaken for outer route")
    require(oracle._tsmis_description(
        oracle.NULL_DESCRIPTION, "001", pdf=True)[0] == "",
        "PDF null Description did not project to blank")

    left = [
        row("left-a", county="AA", description="BETA"),
        row("left-b", county="BB", description="ALPHA"),
    ]
    right = [
        row("right-a", county="AA", description="ALPHA"),
        row("right-b", county="BB", description="BETA"),
    ]
    compared = oracle._compare(
        "county swap", left, right, ("description",),
        collapse_description=False)
    require(compared["counts"]["differing_rows"] == 2,
            "approved D4 county identity masked a cross-county value swap")
    weak = oracle._weak_identity_census([*left, *right])
    require(weak["cross_county_weak_key_count"] == 1
            and weak["cross_county_identity_count"] == 2,
            "weak Route+PM collision census drift")

    district = oracle._compare(
        "district", [row("d12", district="12")], [row("d11", district="11")],
        ("district",), collapse_description=False)
    require(district["counts"]["per_field"]["district"] == 1,
            "District difference was not asserted")
    pr = oracle._compare(
        "pr", [row("pr-r", pr="R")], [row("pr-s", pr="S")],
        ("pr",), collapse_description=False)
    require(pr["counts"]["paired_rows"] == 1
            and pr["counts"]["per_field"]["pr"] == 1,
            "PR incorrectly entered the key or escaped assertion")
    suffix_rows = [row("suffix", pm_suffix="L", hg="L")]
    suffix_contract = oracle._source_claim_contract(suffix_rows)
    require(suffix_contract["pm_suffix_nonblank"] == 1
            and suffix_contract["all_nonblank_pm_suffix_equals_hg"],
            "PM_SFX source claim was not independently conserved")

    duplicate_left = [
        row("a", description="WRONG"), row("b", description="RIGHT")]
    duplicate_right = [row("z", description="RIGHT")]
    pairs, extras, _right_extra, trace = oracle._pair_group(
        duplicate_left, duplicate_right, ("description",),
        collapse_description=False)
    require(pairs[0][0].member == "b.xlsx" and extras[0].member == "a.xlsx"
            and trace["cost"] == 0,
            "exact duplicate assignment did not choose minimum asserted cost")
    rejects(lambda: oracle._pair_group(
        [row(f"l{i}") for i in range(9)], [row("r")], ("description",),
        collapse_description=False), "over-cap duplicate group was accepted")

    require(oracle._pdf_header([word("CITY", 10)]) is None,
            "data word CITY was mistaken for a repeated header")
    require(oracle._pdf_header(valid_header_words()) is not None,
            "complete PDF header constellation was rejected")
    rejects(lambda: oracle._pdf_header(
        [*valid_header_words(10), *valid_header_words(30)]),
        "two complete PDF header constellations were accepted")
    require("Q" not in oracle.PDF_PREFIXES and set("CDGHLMNRST") == oracle.PDF_PREFIXES,
            "PDF postmile-prefix vocabulary drift")

    pdf_rows = [
        row("p1", source="TSMIS PDF", pm="1.000", description="A B",
            raw_description="001/A B"),
        row("p2", source="TSMIS PDF", pm="2.000", description="EDGE",
            raw_description="001/EDGE"),
        row("p3", source="TSMIS PDF", route="010", pm="3.000",
            description="REST", raw_description="010/REST"),
        row("p4", source="TSMIS PDF", pm="4.000", description="",
            area4="", raw_description=oracle.NULL_DESCRIPTION, raw_area4="-"),
    ]
    excel_rows = [
        row("e1", source="TSMIS Excel", pm="1.000", description="A  B",
            raw_description="001/A  B"),
        row("e2", source="TSMIS Excel", pm="2.000", description="EDGE",
            raw_description="001/EDGE \n"),
        row("e3", source="TSMIS Excel", route="010", pm="3.000",
            description="REST_x000d_", raw_description="010/REST_x000d_\n"),
        row("e4", source="TSMIS Excel", pm="4.000", description="",
            area4="", raw_description="", raw_area4=""),
    ]
    render = oracle._classify_tsmis_render_equivalence(excel_rows, pdf_rows)
    require(render["counts"] == {
        "excel_literal_x000d_escape_absent_from_pdf": 1,
        "html_whitespace_collapse": 2,
        "pdf_dash_renders_excel_blank": 1,
        "pdf_no_linear_event_renders_excel_blank": 1,
    } and render["all_classified"], "known raw-render classes drifted")
    require(render["whitespace_equivalent_rows"] == 3
            and render["whitespace_breakdown"] == {
                "edge_only": 2, "internal_runs_or_internal_plus_edge": 1},
            "compound whitespace/escape classification drifted")

    with tempfile.TemporaryDirectory(prefix="rd-stage8-gate-") as temporary:
        root = Path(temporary)
        normalized = root / "normalized.xlsx"
        write_normalized(normalized)
        parsed = oracle._parse_tsn_normalized(normalized)
        require(parsed["rows"] == 1 and parsed["rows_data"][0].pm == "1.200",
                "valid normalized workbook did not parse exactly")
        bad_header = root / "bad-header.xlsx"
        write_normalized(bad_header, mutate=lambda _w, ws: setattr(
            ws["A1"], "value", "Wrong"))
        rejects(lambda: oracle._parse_tsn_normalized(bad_header),
                "normalized header mutation was accepted")
        extra_width = root / "extra-width.xlsx"
        write_normalized(extra_width, mutate=lambda _w, ws: setattr(
            ws["P1"], "value", "Extra"))
        rejects(lambda: oracle._parse_tsn_normalized(extra_width),
                "normalized physical-width mutation was accepted")
        formula = root / "formula.xlsx"
        write_normalized(formula, mutate=lambda _w, ws: setattr(
            ws["I2"], "value", "=1+1"))
        rejects(lambda: oracle._parse_tsn_normalized(formula),
                "normalized formula mutation was accepted")

        tsmis_root = root / "tsmis"
        tsmis_root.mkdir()
        valid_tsmis = tsmis_root / "tsar_ramp_detail_route_001.xlsx"
        write_tsmis(valid_tsmis)
        parsed_tsmis = oracle._parse_tsmis_xlsx(tsmis_root)
        require(parsed_tsmis["rows"] == 1
                and parsed_tsmis["rows_data"][0].description == "DESC",
                "valid TSMIS workbook did not parse exactly")
        product_consolidated = root / "product-consolidated.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = oracle.TSMIS_SHEET
        worksheet.append(["Route", *oracle.TSMIS_HEADER])
        source_values = [
            "07-LA-001", "R", "1.2", date(2020, 1, 2), None, "D", "N",
            "LA", "U", "001/DESC", None,
        ]
        worksheet.append(["001", *source_values])
        workbook.save(product_consolidated)
        workbook.close()
        truth_row = oracle._row_from_tsmis_values(
            source="truth", member="truth.xlsx", source_record="row 2",
            file_route="001", values=source_values, pdf=False)
        inspected = oracle._inspect_product_consolidated(
            product_consolidated, pdf=False, truth_rows=[truth_row])
        require(inspected["projection_exact"],
                "physically omitted declared trailing blank was misread as drift")
        valid_tsmis.unlink()
        wrong_route = tsmis_root / "tsar_ramp_detail_route_002.xlsx"
        write_tsmis(wrong_route)
        rejects(lambda: oracle._parse_tsmis_xlsx(tsmis_root),
                "filename/Location route conflict was accepted")

        zip_a = root / "a.xlsx"
        zip_b = root / "b.xlsx"
        zip_c = root / "c.xlsx"
        for path, core, sheet in (
                (zip_a, b"time-a", b"same"),
                (zip_b, b"time-b", b"same"),
                (zip_c, b"time-a", b"changed")):
            with zipfile.ZipFile(path, "w") as archive:
                archive.writestr("docProps/core.xml", core)
                archive.writestr("xl/worksheets/sheet1.xml", sheet)
        require(oracle._zip_digest_without_core(zip_a)
                == oracle._zip_digest_without_core(zip_b),
                "core.xml-only volatility changed stable package digest")
        require(oracle._zip_digest_without_core(zip_a)
                != oracle._zip_digest_without_core(zip_c),
                "worksheet mutation escaped stable package digest")

        output = root / "result.json"
        oracle._atomic_write_text(output, "one\n")
        decision_path = root / "result.json.acceptance.json"
        decision = oracle._write_decision(
            decision_path, output, {
                "audit": "synthetic", "source_truth_exact": True,
                "production_tsmis_projection_exact": True,
                "production_value_projection_exact": False,
                "production_comparison_semantics_exact": False,
                "stage8_base_oracle_complete": True,
                "comparison_end_to_end_perfect": False,
            }, accepted=True, reason="synthetic",
            postwrite_current=True, postwrite_detail={},
            open_findings_authorized=True)
        require(decision["result_sha256"] == oracle._sha_file(output)
                and decision_path.is_file(),
                "detached acceptance did not bind the exact result")
        oracle._atomic_write_text(output, "two\n")
        require(decision["result_sha256"] != oracle._sha_file(output),
                "detached acceptance followed a later result mutation")

    print(
        f"PASS ({passed} assertions): D4 identity; District/PR/PM_SFX claims; "
        "conditional Description prefixes; exact duplicate assignment; PM/date "
        "domains; PDF header/prefix gates; render classes; normalized/TSMIS XLSX "
        "layout/formula rejection; package volatility; detached acceptance")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
