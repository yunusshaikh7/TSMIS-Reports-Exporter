#!/usr/bin/env python3
"""Run one product leg against the direct-source Highway Sequence raw-TSN twin.

This audit runner is intentionally one-shot and non-accepting.  Before product
code can be imported it authenticates the four-file direct twin, the exact
TSMIS inputs, the accepted Stage-6 chain, and the raw TSN PDF universe.  It
also proves that the requested output tree is disjoint from every bound input
tree in both directions.  A successful leg is evidence for later family
certification; it is never, by itself, Stage-8 family acceptance.
"""

from __future__ import annotations

import argparse
from collections import Counter
import contextlib
from dataclasses import dataclass
import hashlib
import io
import importlib.metadata
import json
import os
from pathlib import Path
import stat
import sys
import tempfile
from types import SimpleNamespace
from typing import Callable, Iterable, Mapping, Sequence
from zipfile import ZipFile
from xml.etree import ElementTree

from openpyxl import load_workbook

import run_phase8_highway_sequence_product_comparison_leg as witness
import run_phase8_highway_sequence_product_raw_tsn_leg as dev


_REPORTED_RUNNER_PATH = Path(__file__)
RUNNER_LEXICAL_PATH = (
    _REPORTED_RUNNER_PATH
    if _REPORTED_RUNNER_PATH.is_absolute()
    else Path.cwd() / _REPORTED_RUNNER_PATH
)
REPO_ROOT = RUNNER_LEXICAL_PATH.parent.parent
BUILD_ROOT = REPO_ROOT / "build"
SCRIPTS_ROOT = REPO_ROOT / "scripts"
VISUAL_ROOT = witness.VISUAL_ROOT
SOURCE_ROOT = VISUAL_ROOT / "phase8_highway_sequence_product_sources_r2"
EXCEL_INPUT = SOURCE_ROOT / "current_tsmis_excel_consolidated.xlsx"
PDF_INPUT = SOURCE_ROOT / "current_tsmis_pdf_consolidated.xlsx"
DIRECT_BUILDER = BUILD_ROOT / "build_phase8_highway_sequence_raw_tsn_direct_twin.py"

LEG_CHOICES = ("excel_vs_raw_tsn", "pdf_vs_raw_tsn")
WORKBOOK_NAME = "highway_sequence_raw_tsn_audit_twin.xlsx"
PROVENANCE_NAME = "highway_sequence_raw_tsn_audit_twin.provenance.json"
MANIFEST_NAME = "manifest.json"
TWIN_RESULT_NAME = "result.json"
TWIN_OUTPUT_NAMES = (
    MANIFEST_NAME,
    PROVENANCE_NAME,
    TWIN_RESULT_NAME,
    WORKBOOK_NAME,
)
EXPECTED_DIRECT_TWIN_ARTIFACTS = {
    WORKBOOK_NAME: (
        2_422_010,
        "68b28921c4ca8290810c92653b4a96077d6a28bdb7954447c287cf3e78d3f67d",
    ),
    PROVENANCE_NAME: (
        31_368_272,
        "95c0229fc0c96eb2f1e8966c300c5916c0978a17f73c39cdf829f909a1ff441b",
    ),
    MANIFEST_NAME: (
        388_864,
        "97541aaa963d784dbf6537cf3e6f46d32fb161f012be0ecb3abda441708b1d91",
    ),
    TWIN_RESULT_NAME: (
        5_183,
        "d4c0a5759b0ca9731047b0f7d57fabedb228f7a61697f0c6af3cb4ef8fc4d134",
    ),
}
SHEET_NAME = "Highway Locations (TSN)"
HEADERS = (
    "Route", "County", "PM", "City", "HG", "FT",
    "Distance To Next Point", "Description",
)
EXPECTED_ORDERED_ROWS_SHA256 = (
    "5ef81b31622730e8f1369d1989cc92c717be7eb4ad8f29061b3750ff78f767fc"
)
EXPECTED_COUNTS = {
    "raw_records": 69_804,
    "data_records": 68_806,
    "equate_records": 998,
    "blank_county_equates": 46,
    "projectable_records": 69_758,
    "pointer_P": 283,
    "pointer_arrow": 282,
    "pointer_total": 565,
    "ordered_rows_sha256": EXPECTED_ORDERED_ROWS_SHA256,
}
EXPECTED_TSMIS = {
    "current_excel": (
        EXCEL_INPUT,
        2_424_212,
        "cf5905332db3d3eb5a49a87d603f6e36f209cad9a84173b381dace6600168b20",
    ),
    "current_pdf": (
        PDF_INPUT,
        2_371_547,
        "070afe51ea3bf84c9704d0a36a02702b65189941badab6374b03461db8ef6ccc",
    ),
}
EXPECTED_SOURCE_TREE_LEDGER_SHA256 = (
    "5c8d3f52e24c92f65beba4ca58d75cae51794e36c8b725df2e811714caff00af"
)
EXPECTED_STATIC_BINDINGS = {
    "accepted_normalized_tsn": (
        2_536_901,
        "9dc84c661a9284131baf928767e210a6d708c0a338819fca2b69b907f85dd041",
    ),
    "accepted_stage6_decision": (
        5_934,
        "71fe59a5f4676d3b935bcbea380374b14fdccfd77b674ea88148fa18760ffde2",
    ),
    "accepted_stage6_result": (
        1_276_684,
        "bdd344258ced0e138196c518be2d49ee058f5f9c0f52dea860c328fc3216d1e2",
    ),
    "stage6_oracle": (
        63_233,
        "0d6cacfa5a4615a80381b077780b051127958bbf325979cf24b7a5c29eb8e17b",
    ),
    "xlsx_reader": (
        40_888,
        "bbfda5ccdbea3697978c0ba4414b7dccf3d5c248ba6762aa946c76e920fc940b",
    ),
}

PINNED_BUILDER_IDENTITY: tuple[int, str] | None = (
    57_219,
    "86d271619f4e446590fe6edaa40e9e85d74da2ca9623f9a5bfcf7877c7101ea5",
)
AUDIT_CODE_LOGICAL_PATHS = {
    "direct_product_runner": "build/run_phase8_highway_sequence_product_direct_raw_tsn_leg.py",
    "shared_product_witness": "build/run_phase8_highway_sequence_product_comparison_leg.py",
    "raw_chunk_witness": "build/run_phase8_highway_sequence_product_raw_tsn_leg.py",
    "direct_twin_builder": "build/build_phase8_highway_sequence_raw_tsn_direct_twin.py",
}

PRODUCT_MODULES = {
    "compare_highway_sequence_tsn",
    "compare_highway_sequence_pdf",
    "compare_core",
    "consolidation_meta",
}
RESULT_NAME = witness.RESULT_NAME
ARTIFACT_MANIFEST_NAME = witness.ARTIFACT_MANIFEST_NAME
PRODUCT_CODE_MANIFEST_NAME = witness.PRODUCT_CODE_MANIFEST_NAME
COMPLETION_NAME = "completion.json"
REPARSE_FLAG = int(getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400))
ARTIFACT_STATUS = "DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE"
PRETERMINAL_AUDIT = (
    "Stage 8 Highway Sequence direct-source raw-TSN product comparison "
    "leg preterminal record"
)
PRETERMINAL_STATUS = (
    "PRETERMINAL_AUDIT_WITNESS_PENDING_DETACHED_COMPLETION"
)
COMPLETION_AUDIT = (
    "Stage 8 Highway Sequence direct-source raw-TSN product comparison "
    "leg detached completion"
)
COMPLETION_STATUS = "PASS_DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE"

PRETERMINAL_KEYS = {
    "acceptance_eligible", "artifact_manifest", "artifact_manifest_before_result",
    "artifact_status", "audit", "audit_code", "audit_code_mutations",
    "decoded_comparison_payload", "deterministic_serialization",
    "direct_twin_preimport_validation", "expected_precompletion_artifact_names",
    "input_tree_revalidation", "invariants", "leg", "loaded_product_code",
    "outcome_sidecars", "output_containment_mutations", "output_root", "outputs",
    "product_code_manifest", "publication_artifacts",
    "publication_lifecycle_mutations", "reason", "required_detached_terminal_completion",
    "residue_gate", "result", "schema_version", "stage8_family_accepted",
    "status", "terminal",
}
PRETERMINAL_INVARIANT_KEYS = {
    "accepted_stage6_and_raw_pdf_bindings_exact", "bidirectional_998_998_zero_orphan_topology",
    "canonical_deterministic_audit_json", "committed_formula_value_twin",
    "complete_ok_zero_zero", "direct_twin_current_builder_exact",
    "direct_twin_not_family_acceptance", "direct_twin_v1_validated_before_product_import",
    "disposable_containment_mutations_passed", "exact_artifact_universe_declared",
    "input_tree_universes_frozen_through_pre_result", "loaded_product_code_manifested",
    "no_delete_or_overwrite", "no_transient_residue", "one_leg_only", "pairing_exact",
    "payload_chunks_decoded_and_bound", "publication_lifecycle_mutations_passed",
    "raw_records_69804", "referenced_decoded_inventoried_chunks_equal",
    "reverse_only_topology_mutation_rejected", "this_record_is_explicitly_nonterminal",
    "tsmis_inputs_independently_exact_bound", "two_trusted_outcome_sidecars",
    "two_way_output_input_disjointness", "only_zero_byte_source_backed_permanent_lease",
    "workbook_reopen_equals_all_provenance_rows",
}
COMPLETION_KEYS = {
    "acceptance_eligible", "artifact_status", "audit", "audit_code",
    "audit_code_mutations", "complete_output_artifact_manifest",
    "expected_final_artifact_names", "invariants", "leg",
    "output_containment_mutations", "output_root", "payload_chunk_names",
    "post_result_input_revalidation", "preterminal_result",
    "publication_lifecycle_mutations", "schema_version", "stage8_family_accepted",
    "status", "terminal", "terminal_preconditions", "terminal_residue_gate",
}
COMPLETION_INVARIANT_KEYS = {
    "complete_output_artifact_universe_exact", "complete_output_identities_exact",
    "complete_output_members_physically_disjoint", "nonterminal_result_exactly_bound",
    "not_family_acceptance", "output_root_still_disjoint",
    "post_result_input_trees_exact", "referenced_decoded_inventoried_final_chunks_equal",
    "terminal_completion_is_last_commit", "terminal_residue_exact",
}
AUDIT_CODE_MANIFEST_KEYS = {
    "canonical_members_sha256", "members", "roles", "schema",
}
AUDIT_CODE_MEMBER_KEYS = {
    "bytes", "logical_path", "role", "sha256",
}
COMPLETE_OUTPUT_MANIFEST_KEYS = {
    "bytes", "canonical_members_sha256", "files", "members", "schema", "scope",
}
COMPLETE_OUTPUT_MEMBER_KEYS = {"bytes", "relative_path", "sha256"}
IDENTITY_KEYS = {"bytes", "path", "sha256"}
AUDIT_CODE_MUTATION_KEYS = {
    "every_copied_role_byte_mutation_rejected", "mutated_roles",
    "every_reported_role_reparse_redirection_rejected",
    "physical_logical_paths_exact", "real_audit_code_unchanged", "roles_copied",
    "reported_role_reparse_redirections_rejected",
    "unchanged_copied_baseline_passed",
}
CONTAINMENT_MUTATION_KEYS = {
    "broken_link_occupant_mutation", "child_of_input_rejected",
    "directory_link_mutation", "file_link_mutation", "hardlink_alias_rejected",
    "lexical_alias_rejected", "parent_of_input_rejected", "same_root_rejected",
    "valid_sibling_accepted",
}
INPUT_REVALIDATION_KEYS = {
    "bound_files_exact", "phase", "tree_universes_exact", "trees",
}
RESIDUE_KEYS = {
    "exact_artifact_universe", "permanent_lease", "permanent_lease_exception",
    "rejected_name_classes", "transient_residue",
}
RESIDUE_UNIVERSE_KEYS = {"audit_names", "core_names", "payload_chunks"}
AUDIT_CODE_SEMANTIC_MUTATION_LABELS = {
    f"audit_code_role_{role}" for role in AUDIT_CODE_LOGICAL_PATHS
}
COMPLETION_SEMANTIC_MUTATION_LABELS = {
    "acceptance_true", "artifact_status_drift", "audit_code_binding", "audit_drift",
    "audit_mutation_evidence", "containment_evidence", "expected_final_names",
    "extra_top_level", "family_acceptance_true", "invariant_false",
    "manifest_aggregate", "manifest_self_inclusion", "missing_audit",
    "output_root_drift", "payload_chunk_universe", "post_result_phase",
    "post_result_tree", "preterminal_result_identity", "publication_controls_extra_key",
    "publication_truth_false", "schema_version_drift", "status_drift",
    "terminal_false", "terminal_precondition_false", "terminal_residue",
    "wrong_valid_leg",
} | AUDIT_CODE_SEMANTIC_MUTATION_LABELS
PUBLICATION_MUTATION_KEYS = {
    "all_failed_preconditions_left_no_terminal_file", "all_publication_controls_passed",
    "bound_input_hardlink_rejected", "completion_semantic_mutations_left_no_terminal_file",
    "completion_semantic_mutations_rejected", "completion_contract_baseline_accepted",
    "destination_collision_rejected_without_overwrite",
    "disposable_control_only", "extra_name_universe_rejected", "failed_terminal_preconditions",
    "final_artifact_symlink_mutation", "missing_name_universe_rejected",
    "pairwise_output_hardlink_rejected", "post_manifest_artifact_roles_rejected",
    "second_callback_artifact_mutation_rejected", "staging_payload_mutation_rejected",
    "staging_residue_rejected", "successful_terminal_control",
    "unchanged_baseline_accepted",
}
SUCCESSFUL_TERMINAL_CONTROL_KEYS = {
    "completion_canonical_and_exact", "exact_final_universe", "exit_zero",
    "no_pending_residue", "precommit_stdout_nonterminal", "validation_phases",
}
PRODUCT_RESULT_KEYS = {
    "artifact_generation", "comparison_outcome_sha256", "completion", "counts",
    "duplicate_group_count", "failed_inputs", "failures", "pairing_quality",
    "pairing_trace_count", "pairing_trace_sha256", "persisted_members",
    "skipped_inputs", "status", "summary_lines", "verdict", "warnings",
}
PRODUCT_COUNT_KEYS = {
    "asserted_cells", "context_cells", "differing_cells", "differing_rows", "known",
    "paired_rows", "per_field_counts", "side_a_only_rows", "side_b_only_rows",
}
ARTIFACT_GENERATION_KEYS = {"completion", "members", "publication_state", "requested_mode"}
GENERATION_MEMBER_KEYS = {"bytes", "commit_role", "flavor", "path", "sha256"}
PERSISTED_MEMBER_KEYS = {"completion", "current", "source", "trusted"}
DECODED_PAYLOAD_KEYS = {
    "chunks", "comparison_schema_version", "decoded_bytes", "decoded_sha256",
    "encoding", "independent_decode_exact", "payload_schema_version",
    "product_reader_decoded_both_peers",
}
DECODED_CHUNK_KEYS = {
    "compressed_bytes", "compressed_sha256", "decoded_bytes", "decoded_sha256",
    "index", "relative_path",
}
PUBLICATION_ARTIFACT_KEYS = {"outcome_sidecars", "payload_chunks", "permanent_lease"}

DIRECT_RESULT_KEYS = {
    "acceptance_eligible", "artifact_status", "artifact_universe", "artifacts",
    "audit", "bidirectional_equate_topology", "counts", "invariants",
    "ordered_rows_sha256", "output_root_embedded", "reason",
    "reverse_topology_mutation_rejected", "schema_version",
    "stage8_family_accepted", "status", "terminal",
}
DIRECT_MANIFEST_KEYS = {
    "acceptance_eligible", "accepted_stage6_chain", "artifact_status", "audit",
    "bidirectional_equate_topology", "builder_identity", "canonical_xlsx_package",
    "core_timestamp_mutation", "counts", "delayed_in_memory_determinism",
    "fresh_stage6_digests", "generated", "independence",
    "output_artifact_names", "output_root_embedded", "output_root_gate",
    "reverse_topology_mutation", "runtime", "schema_version", "source_capture",
    "stage8_family_accepted", "static_bindings", "workbook_reopen",
    "workbook_schema",
}
DIRECT_PROVENANCE_KEYS = {
    "acceptance_eligible", "accepted_stage6_chain", "artifact_status", "audit",
    "bidirectional_equate_topology", "raw_documents", "reason", "row_count", "rows",
    "schema", "schema_version", "source_capture", "stage8_family_accepted",
}
DIRECT_COUNT_KEYS = {
    "blank_county_equates", "data_records", "equate_records",
    "ordered_rows_sha256", "pointer_P", "pointer_arrow", "pointer_total",
    "projectable_records", "raw_records", "route_counts",
}
DIRECT_INVARIANT_KEYS = {
    "accepted_stage6_chain_exact", "blank_county_equates_46",
    "canonical_xlsx_bytes", "core_created_modified_exact_fixed",
    "data_records_68806", "delayed_in_memory_xlsx_byte_identical",
    "equate_records_998", "forward_equates_to_e_exact_998",
    "fresh_stage6_digests_exact", "no_output_root_volatility",
    "not_family_acceptance", "output_root_disjointness_and_alias_controls",
    "parser_consumed_exact_captured_payloads",
    "planted_core_time_mutation_recanonicalized_exact",
    "planted_core_time_mutation_rejected", "raw_pointer_tokens_565",
    "raw_records_69804", "render_surface_probe_exact",
    "reverse_e_to_equates_exact_998", "reverse_only_mutation_rejected",
    "source_file_universe_exact",
    "twelve_source_payloads_single_captured_and_exact_bound",
    "xlsx_no_formulas_or_errors", "xlsx_stream_reopen_exact",
    "zero_unpaired_both_directions",
}
STAGE6_TRACKED_KEYS = {
    "family_gate", "lifecycle", "normalized", "oracle", "reader",
    "reader_gate", "result", "sidecar",
}
FRESH_STAGE6_KEYS = {
    "all_equal_accepted_stage6", "document_metadata_digests",
    "raw_source_digests", "raw_source_provenance_digests",
}


class DirectRawWitnessError(witness.WitnessError):
    """A direct-twin, containment, or product witness contract failed."""


@dataclass(frozen=True)
class CapturedFile:
    path: Path
    payload: bytes
    bytes: int
    sha256: str
    token: tuple[object, ...]
    nlink: int

    def identity(self) -> dict[str, object]:
        return {
            "path": str(self.path),
            "bytes": self.bytes,
            "sha256": self.sha256,
        }


@dataclass(frozen=True)
class TreeSnapshot:
    root: Path
    root_token: tuple[object, ...]
    entries: tuple[tuple[object, ...], ...]

    def public(self) -> dict[str, object]:
        members = [
            {
                "relative_path": item[0],
                "kind": item[1],
                "bytes": item[2],
                "sha256": item[3],
            }
            for item in self.entries
        ]
        return {
            "root": str(self.root),
            "entries": len(members),
            "files": sum(item["kind"] == "file" for item in members),
            "directories": sum(item["kind"] == "directory" for item in members),
            "bytes": sum(int(item["bytes"]) for item in members),
            "canonical_members_sha256": hashlib.sha256(
                witness._canonical_bytes(members)
            ).hexdigest(),
        }


@dataclass(frozen=True)
class Preflight:
    summary: dict[str, object]
    bound_files: tuple[Path, ...]
    protected_roots: tuple[Path, ...]
    tree_snapshots: tuple[TreeSnapshot, ...]
    bound_identities: tuple[tuple[str, int, str, tuple[object, ...]], ...]
    raw_tsn_workbook: Path


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise DirectRawWitnessError(message)


def _reject_constant(value: str) -> None:
    raise DirectRawWitnessError(f"non-finite JSON constant is forbidden: {value}")


def _unique_object(pairs: list[tuple[str, object]]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in pairs:
        if key in result:
            raise DirectRawWitnessError(
                f"duplicate JSON object key is forbidden: {key!r}"
            )
        result[key] = value
    return result


def _strict_json_bytes(raw: bytes, label: str, *, canonical: bool) -> object:
    try:
        value = json.loads(
            raw.decode("utf-8"),
            object_pairs_hook=_unique_object,
            parse_constant=_reject_constant,
        )
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise DirectRawWitnessError(
            f"{label} is not strict UTF-8 JSON: {exc}"
        ) from exc
    if canonical and raw != witness._canonical_bytes(value, newline=True):
        raise DirectRawWitnessError(f"{label} is not canonical LF-terminated JSON")
    return value


def _stat_token(value: os.stat_result) -> tuple[object, ...]:
    return (
        value.st_dev,
        value.st_ino,
        stat.S_IFMT(value.st_mode),
        value.st_size,
        value.st_mtime_ns,
        getattr(value, "st_ctime_ns", None),
        value.st_nlink,
        int(getattr(value, "st_file_attributes", 0)),
    )


def _directory_token(value: os.stat_result | SimpleNamespace) -> tuple[object, ...]:
    """Stable directory identity without Windows' lazy ``st_size`` hydration.

    File size remains mandatory in ``_stat_token``. Directory membership and
    every child file byte are independently inventoried by ``_snapshot_tree``;
    the directory allocation-size hint is neither content nor stable before
    versus after its first Windows enumeration.
    """
    return (
        value.st_dev,
        value.st_ino,
        stat.S_IFMT(value.st_mode),
        value.st_mtime_ns,
        getattr(value, "st_ctime_ns", None),
        value.st_nlink,
        int(getattr(value, "st_file_attributes", 0)),
    )


def _run_directory_token_mutation_probes() -> dict[str, object]:
    facts = os.lstat(VISUAL_ROOT)
    fields = {
        "st_dev": facts.st_dev,
        "st_ino": facts.st_ino,
        "st_mode": facts.st_mode,
        "st_size": facts.st_size,
        "st_mtime_ns": facts.st_mtime_ns,
        "st_ctime_ns": getattr(facts, "st_ctime_ns", None),
        "st_nlink": facts.st_nlink,
        "st_file_attributes": int(getattr(facts, "st_file_attributes", 0)),
    }
    size_only = dict(fields)
    size_only["st_size"] = int(fields["st_size"]) + 4096
    planted_size = SimpleNamespace(**size_only)
    _require(
        _directory_token(facts) == _directory_token(planted_size),
        "directory-size-only mutation changed the stable directory token",
    )
    _require(
        _stat_token(facts) != _stat_token(planted_size),
        "full ordinary-file token stopped binding size",
    )

    mutations = {
        "st_dev": int(fields["st_dev"]) + 1,
        "st_ino": int(fields["st_ino"]) + 1,
        "st_mode": stat.S_IFREG | (int(fields["st_mode"]) & 0o777),
        "st_mtime_ns": int(fields["st_mtime_ns"]) + 1,
        "st_ctime_ns": int(fields["st_ctime_ns"] or 0) + 1,
        "st_nlink": int(fields["st_nlink"]) + 1,
        "st_file_attributes": int(fields["st_file_attributes"]) ^ REPARSE_FLAG,
    }
    rejected_fields: list[str] = []
    baseline = _directory_token(facts)
    for field, value in mutations.items():
        planted_fields = dict(fields)
        planted_fields[field] = value
        planted = SimpleNamespace(**planted_fields)
        _require(
            _directory_token(planted) != baseline,
            f"retained directory-token field mutation escaped: {field}",
        )
        rejected_fields.append(field)

    with tempfile.TemporaryDirectory(
        prefix="phase8-direct-directory-token-control-", dir=VISUAL_ROOT,
    ) as temporary:
        root = Path(temporary).resolve(strict=True)
        member = root / "member.bin"
        baseline_payload = b"directory-token-member-baseline"
        member.write_bytes(baseline_payload)
        baseline_snapshot = _snapshot_tree(root)

        added = root / "added.bin"
        added.write_bytes(b"added-member")
        added_snapshot = _snapshot_tree(root)
        _require(
            added_snapshot != baseline_snapshot,
            "added tree member escaped directory snapshot",
        )
        added.unlink()

        member.write_bytes(baseline_payload + b"-changed")
        changed_snapshot = _snapshot_tree(root)
        _require(
            changed_snapshot != baseline_snapshot,
            "changed tree member escaped directory snapshot",
        )
        member.write_bytes(baseline_payload)

        member.unlink()
        removed_snapshot = _snapshot_tree(root)
        _require(
            removed_snapshot != baseline_snapshot,
            "removed tree member escaped directory snapshot",
        )

    return {
        "directory_size_only_mutation_ignored": True,
        "ordinary_file_size_remains_bound": True,
        "retained_directory_fields_mutated_and_rejected": rejected_fields,
        "all_retained_directory_fields_bound": set(rejected_fields) == set(mutations),
        "added_member_rejected": True,
        "changed_member_rejected": True,
        "removed_member_rejected": True,
    }


def _lexical_absolute(path: Path) -> Path:
    _require(path.is_absolute(), f"path must be absolute: {path}")
    _require(
        not any(part in {".", ".."} for part in path.parts),
        f"lexical alias component is forbidden: {path}",
    )
    return path


def _lexical_abspath(path: Path) -> Path:
    """Make a reported path absolute without resolving or normalizing aliases."""
    return _lexical_absolute(path if path.is_absolute() else Path.cwd() / path)


def _existing_components(
    path: Path,
) -> list[tuple[Path, os.stat_result]]:
    path = _lexical_absolute(path)
    anchor = Path(path.anchor)
    components: list[tuple[Path, os.stat_result]] = []
    current = anchor
    candidates = [anchor]
    for part in path.parts[1:]:
        current = current / part
        candidates.append(current)
    for candidate in candidates:
        try:
            facts = os.lstat(candidate)
        except FileNotFoundError:
            break
        except OSError as exc:
            raise DirectRawWitnessError(
                f"cannot lstat supplied lexical component: {candidate}"
            ) from exc
        components.append((candidate, facts))
    return components


def _assert_plain_components(path: Path, *, include_leaf: bool = True) -> None:
    inspected = _existing_components(path)
    if not include_leaf and inspected and inspected[-1][0] == path:
        inspected = inspected[:-1]
    for component, facts in inspected:
        if stat.S_ISLNK(facts.st_mode) or (
            int(getattr(facts, "st_file_attributes", 0)) & REPARSE_FLAG
        ):
            raise DirectRawWitnessError(
                f"reparse/symlink path component is forbidden: {component}"
            )


def _capture_file(path: Path) -> CapturedFile:
    path = _lexical_absolute(path)
    _assert_plain_components(path)
    try:
        before = os.lstat(path)
    except OSError as exc:
        raise DirectRawWitnessError(f"required file is absent: {path}") from exc
    _require(stat.S_ISREG(before.st_mode), f"not an ordinary file: {path}")
    _require(
        not (int(getattr(before, "st_file_attributes", 0)) & REPARSE_FLAG),
        f"reparse file is forbidden: {path}",
    )
    try:
        payload = path.read_bytes()
        after = os.lstat(path)
        resolved = path.resolve(strict=True)
    except (OSError, RuntimeError) as exc:
        raise DirectRawWitnessError(f"cannot capture required file: {path}") from exc
    _require(
        _stat_token(before) == _stat_token(after),
        f"file changed during exact-byte capture: {path}",
    )
    return CapturedFile(
        path=resolved,
        payload=payload,
        bytes=len(payload),
        sha256=hashlib.sha256(payload).hexdigest(),
        token=_stat_token(after),
        nlink=int(after.st_nlink),
    )


def _identity_tuple(captured: CapturedFile) -> tuple[str, int, str]:
    return str(captured.path), captured.bytes, captured.sha256


def _frozen_identity_tuple(
    captured: CapturedFile,
) -> tuple[str, int, str, tuple[object, ...]]:
    return str(captured.path), captured.bytes, captured.sha256, captured.token


def _embedded_identity_matches(
    embedded: object,
    captured: CapturedFile,
) -> bool:
    if not isinstance(embedded, Mapping):
        return False
    raw_path = embedded.get("canonical_path", embedded.get("path"))
    if raw_path is None:
        path_matches = True
    else:
        try:
            path_matches = Path(str(raw_path)).resolve(strict=True) == captured.path
        except (OSError, RuntimeError, ValueError):
            return False
    size = embedded.get("bytes", embedded.get("size"))
    return (
        path_matches
        and size == captured.bytes
        and embedded.get("sha256") == captured.sha256
    )


def _artifact_identity_matches(embedded: object, captured: CapturedFile) -> bool:
    return (
        isinstance(embedded, Mapping)
        and embedded.get("bytes") == captured.bytes
        and embedded.get("sha256") == captured.sha256
    )


def _snapshot_tree(root: Path) -> TreeSnapshot:
    root = _lexical_absolute(root)
    _assert_plain_components(root)
    try:
        root_before = os.lstat(root)
    except OSError as exc:
        raise DirectRawWitnessError(f"input tree is absent: {root}") from exc
    _require(stat.S_ISDIR(root_before.st_mode), f"input tree is not a directory: {root}")
    canonical = root.resolve(strict=True)
    entries: list[tuple[object, ...]] = []

    def visit(directory: Path) -> None:
        try:
            children = sorted(
                directory.iterdir(), key=lambda item: item.name
            )
        except OSError as exc:
            raise DirectRawWitnessError(f"cannot enumerate input tree: {directory}") from exc
        for child in children:
            relative = child.relative_to(canonical).as_posix()
            _assert_plain_components(child)
            facts = os.lstat(child)
            _require(
                not stat.S_ISLNK(facts.st_mode)
                and not (
                    int(getattr(facts, "st_file_attributes", 0)) & REPARSE_FLAG
                ),
                f"input tree contains a reparse entry: {child}",
            )
            if stat.S_ISDIR(facts.st_mode):
                entry_index = len(entries)
                entries.append((relative, "directory", 0, None, None))
                visit(child)
                after = os.lstat(child)
                _require(
                    _directory_token(facts) == _directory_token(after),
                    f"input directory changed during snapshot: {child}",
                )
                entries[entry_index] = (
                    relative, "directory", 0, None, _directory_token(after),
                )
            elif stat.S_ISREG(facts.st_mode):
                captured = _capture_file(child)
                entries.append((
                    relative, "file", captured.bytes, captured.sha256,
                    captured.token,
                ))
            else:
                raise DirectRawWitnessError(
                    f"input tree contains a non-file/non-directory: {child}"
                )
    visit(canonical)
    root_after = os.lstat(canonical)
    _require(
        _directory_token(root_before) == _directory_token(root_after),
        f"input tree root changed during snapshot: {root}",
    )
    return TreeSnapshot(
        root=canonical,
        root_token=_directory_token(root_after),
        entries=tuple(entries),
    )


def _paths_overlap(left: Path, right: Path) -> bool:
    left = left.resolve(strict=False)
    right = right.resolve(strict=False)
    return left == right or left in right.parents or right in left.parents


def _same_existing_object(left: Path, right: Path) -> bool:
    try:
        return left.exists() and right.exists() and os.path.samefile(left, right)
    except OSError:
        return False


def _output_policy_violations(
    candidate: Path,
    *,
    protected_files: Sequence[Path],
    protected_roots: Sequence[Path],
    permit_existing_output_root: bool = False,
) -> list[str]:
    violations: list[str] = []
    if not candidate.is_absolute():
        return ["not_absolute"]
    lexical_alias = any(part in {".", ".."} for part in candidate.parts)
    if lexical_alias:
        violations.append("lexical_alias_component")
    if os.path.lexists(candidate) and not permit_existing_output_root:
        violations.append("candidate_exists")
    # CMP-AUD-227 ordering is intentional: inspect every supplied lexical
    # component without following it before any canonical resolution.
    if not lexical_alias:
        try:
            component_facts = _existing_components(candidate)
        except DirectRawWitnessError:
            violations.append("component_lstat_failure")
        else:
            for component, facts in component_facts:
                if stat.S_ISLNK(facts.st_mode) or (
                    int(getattr(facts, "st_file_attributes", 0)) & REPARSE_FLAG
                ):
                    violations.append(f"reparse_component:{component}")
    if lexical_alias or any(
        item.startswith("reparse_component:") for item in violations
    ):
        return sorted(set(violations))
    try:
        canonical = candidate.resolve(strict=False)
        visual = VISUAL_ROOT.resolve(strict=True)
    except (OSError, RuntimeError):
        return sorted(set([*violations, "unresolvable_path"]))
    if canonical == visual or visual not in canonical.parents:
        violations.append("outside_private_visual_root")
    for root in protected_roots:
        if _paths_overlap(canonical, root):
            violations.append(f"overlaps_input_root:{root}")
        if _same_existing_object(candidate, root):
            violations.append(f"aliases_input_root:{root}")
    for path in protected_files:
        if _paths_overlap(canonical, path):
            violations.append(f"overlaps_input_file:{path}")
        if _same_existing_object(candidate, path):
            violations.append(f"aliases_input_file:{path}")
    return sorted(set(violations))


def _validate_output_root(
    candidate: Path,
    *,
    protected_files: Sequence[Path],
    protected_roots: Sequence[Path],
) -> Path:
    violations = _output_policy_violations(
        candidate,
        protected_files=protected_files,
        protected_roots=protected_roots,
        permit_existing_output_root=False,
    )
    _require(not violations, f"unsafe output root: {violations}")
    _require(candidate.parent.is_dir(), f"output parent is absent: {candidate.parent}")
    _assert_plain_components(candidate.parent)
    return candidate.resolve(strict=False)


def _run_disjointness_mutation_probes() -> dict[str, object]:
    VISUAL_ROOT.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix="phase8-direct-raw-containment-", dir=VISUAL_ROOT
    ) as temporary:
        base = Path(temporary).resolve(strict=True)
        input_parent = base / "inputs"
        input_root = input_parent / "bound-tree"
        input_root.mkdir(parents=True)
        input_file = input_root / "bound.bin"
        input_file.write_bytes(b"bound-input")
        protected_files = (input_file.resolve(strict=True),)
        protected_roots = (input_root.resolve(strict=True),)

        def violations(path: Path) -> list[str]:
            return _output_policy_violations(
                path,
                protected_files=protected_files,
                protected_roots=protected_roots,
                permit_existing_output_root=False,
            )

        child = violations(input_root / "forbidden-child")
        parent = violations(input_parent)
        same = violations(input_root)
        lexical = violations(input_parent / ".." / "lexical-output")
        valid = violations(base / "valid-sibling")
        _require(
            any(item.startswith("overlaps_input_root:") for item in child),
            f"child-of-input mutation escaped: {child}",
        )
        _require(
            any(item.startswith("overlaps_input_root:") for item in parent),
            f"parent-of-input mutation escaped: {parent}",
        )
        _require(
            any(item.startswith("overlaps_input_root:") for item in same),
            f"same-root mutation escaped: {same}",
        )
        _require(
            "lexical_alias_component" in lexical,
            f"lexical-alias mutation escaped: {lexical}",
        )
        _require(not valid, f"valid sibling control was rejected: {valid}")

        hardlink = base / "hardlink-output"
        os.link(input_file, hardlink)
        hardlink_violations = violations(hardlink)
        _require(
            any(item.startswith("aliases_input_file:") for item in hardlink_violations),
            f"hardlink-alias mutation escaped: {hardlink_violations}",
        )

        symlink_status = "unexecuted_platform_denied"
        symlink_violations: list[str] = []
        link = base / "directory-link"
        try:
            os.symlink(input_root, link, target_is_directory=True)
        except (OSError, NotImplementedError):
            pass
        else:
            symlink_violations = violations(link / "new-output")
            _require(
                any(item.startswith("reparse_component:") for item in symlink_violations),
                f"directory-link mutation escaped: {symlink_violations}",
            )
            symlink_status = "executed_and_rejected"

        file_link_status = "unexecuted_platform_denied"
        file_link = base / "file-link-output"
        try:
            os.symlink(input_file, file_link, target_is_directory=False)
        except (OSError, NotImplementedError):
            pass
        else:
            file_link_violations = violations(file_link)
            _require(
                "candidate_exists" in file_link_violations
                and any(item.startswith("reparse_component:")
                        for item in file_link_violations),
                f"file-link alias mutation escaped: {file_link_violations}",
            )
            file_link_status = "executed_and_rejected"

        broken_status = "unexecuted_platform_denied"
        broken = base / "broken-link-output"
        try:
            os.symlink(base / "absent-target", broken, target_is_directory=True)
        except (OSError, NotImplementedError):
            pass
        else:
            broken_violations = violations(broken)
            _require(
                "candidate_exists" in broken_violations
                and any(item.startswith("reparse_component:")
                        for item in broken_violations),
                f"broken-link occupant mutation escaped: {broken_violations}",
            )
            broken_status = "executed_and_rejected"

        return {
            "child_of_input_rejected": True,
            "parent_of_input_rejected": True,
            "same_root_rejected": True,
            "lexical_alias_rejected": True,
            "hardlink_alias_rejected": True,
            "directory_link_mutation": symlink_status,
            "file_link_mutation": file_link_status,
            "broken_link_occupant_mutation": broken_status,
            "valid_sibling_accepted": True,
        }


def _require_keys(value: object, expected: set[str], label: str) -> Mapping[str, object]:
    _require(isinstance(value, Mapping), f"{label} must be a JSON object")
    observed = set(value)
    _require(
        observed == expected,
        f"{label} key universe drift: missing={sorted(expected - observed)!r}, "
        f"extra={sorted(observed - expected)!r}",
    )
    return value


def _run_v1_schema_mutation_probes(
    result: Mapping[str, object],
    manifest: Mapping[str, object],
    provenance: Mapping[str, object],
) -> dict[str, object]:
    probes: list[tuple[str, Mapping[str, object], set[str]]] = [
        ("result_top_level", result, DIRECT_RESULT_KEYS),
        ("manifest_top_level", manifest, DIRECT_MANIFEST_KEYS),
        ("provenance_top_level", provenance, DIRECT_PROVENANCE_KEYS),
        ("result_counts", result["counts"], DIRECT_COUNT_KEYS),
        ("result_invariants", result["invariants"], DIRECT_INVARIANT_KEYS),
        (
            "stage6_tracked_identities",
            manifest["accepted_stage6_chain"]["tracked_identities"],
            STAGE6_TRACKED_KEYS,
        ),
        ("fresh_stage6_digests", manifest["fresh_stage6_digests"], FRESH_STAGE6_KEYS),
    ]
    rejected: list[str] = []
    for label, baseline, expected in probes:
        _require_keys(baseline, expected, f"schema baseline {label}")
        mutated = dict(baseline)
        mutated["__unexpected_v1_extension__"] = True
        try:
            _require_keys(mutated, expected, f"schema mutation {label}")
        except DirectRawWitnessError:
            rejected.append(label)
        else:
            raise DirectRawWitnessError(
                f"extra-key v1 schema mutation escaped: {label}"
            )
    return {
        "mutations": len(probes),
        "rejected": rejected,
        "all_extra_key_mutations_rejected": len(rejected) == len(probes),
    }


def _capture_direct_twin_root(
    twin_root: Path,
) -> tuple[Path, dict[str, CapturedFile], TreeSnapshot]:
    twin_root = _lexical_absolute(twin_root)
    _assert_plain_components(twin_root)
    try:
        facts = os.lstat(twin_root)
    except OSError as exc:
        raise DirectRawWitnessError(f"direct twin root is absent: {twin_root}") from exc
    _require(stat.S_ISDIR(facts.st_mode), f"direct twin root is not a directory: {twin_root}")
    canonical = twin_root.resolve(strict=True)
    try:
        entries = sorted(canonical.iterdir(), key=lambda item: item.name)
    except OSError as exc:
        raise DirectRawWitnessError(f"cannot enumerate direct twin root: {canonical}") from exc
    _require(
        [item.name for item in entries] == sorted(TWIN_OUTPUT_NAMES),
        "direct twin does not have the exact four-file universe",
    )
    captures: dict[str, CapturedFile] = {}
    for entry in entries:
        _assert_plain_components(entry)
        entry_facts = os.lstat(entry)
        _require(
            stat.S_ISREG(entry_facts.st_mode),
            f"direct twin contains a nested/non-file entry: {entry}",
        )
        captured = _capture_file(entry)
        _require(
            (captured.bytes, captured.sha256)
            == EXPECTED_DIRECT_TWIN_ARTIFACTS[entry.name],
            f"direct twin final pinned artifact drift: {entry.name}",
        )
        captures[entry.name] = captured
    snapshot = _snapshot_tree(canonical)
    _require(
        len(snapshot.entries) == 4
        and all(item[1] == "file" for item in snapshot.entries),
        "direct twin tree is not exactly four ordinary flat files",
    )
    snapshot_identities = {
        str(item[0]): (int(item[2]), str(item[3])) for item in snapshot.entries
    }
    _require(
        snapshot_identities
        == {name: (item.bytes, item.sha256) for name, item in captures.items()},
        "direct twin changed between exact capture and tree snapshot",
    )
    return canonical, captures, snapshot


def _validate_nonacceptance_contract(
    document: Mapping[str, object],
    *,
    label: str,
    terminal_result: bool = False,
    reason_required: bool = True,
) -> None:
    _require(document.get("schema_version") == 1, f"{label} schema is not v1")
    _require(
        document.get("artifact_status")
        == "DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE",
        f"{label} artifact status drift",
    )
    _require(document.get("acceptance_eligible") is False, f"{label} became acceptance eligible")
    _require(document.get("stage8_family_accepted") is False, f"{label} falsely claims family acceptance")
    if reason_required:
        _require(
            isinstance(document.get("reason"), str) and bool(document.get("reason")),
            f"{label} non-acceptance reason is absent",
        )
    if terminal_result:
        _require(
            document.get("status")
            == "PASS_DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE"
            and document.get("terminal") is True,
            "direct twin result lacks its exact terminal non-acceptance status",
        )


def _ordered_rows_sha256(rows: Iterable[Sequence[object]]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update(json.dumps(
            list(row), ensure_ascii=False, separators=(",", ":"),
        ).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def _public_source_identity(captured: CapturedFile) -> dict[str, object]:
    return {
        "name": captured.path.name,
        "canonical_path": str(captured.path),
        "bytes": captured.bytes,
        "sha256": captured.sha256,
    }


def _bind_embedded_file(
    embedded: object,
    *,
    label: str,
    expected: tuple[int, str] | None = None,
) -> CapturedFile:
    _require(isinstance(embedded, Mapping), f"{label} identity is absent")
    raw_path = embedded.get("canonical_path", embedded.get("path"))
    _require(isinstance(raw_path, str) and raw_path, f"{label} path is absent")
    path = Path(raw_path)
    captured = _capture_file(path)
    _require(
        _embedded_identity_matches(embedded, captured),
        f"{label} embedded identity disagrees with live captured bytes",
    )
    if expected is not None:
        _require(
            (captured.bytes, captured.sha256) == expected,
            f"{label} fixed identity drift: {captured.bytes}/{captured.sha256}",
        )
    return captured


def _validate_source_capture(
    manifest: Mapping[str, object],
    provenance: Mapping[str, object],
) -> tuple[dict[str, CapturedFile], Path, dict[str, object]]:
    source = _require_keys(
        manifest.get("source_capture"),
        {
            "source_members", "source_bytes", "source_identities",
            "source_identity_ledger_sha256", "non_source_role", "file_universe",
            "each_payload_captured_once",
            "parser_consumes_the_same_captured_payload_object",
        },
        "manifest source_capture",
    )
    _require(
        provenance.get("source_capture") == source,
        "manifest/provenance source-capture ledgers disagree",
    )
    identities = source.get("source_identities")
    _require(isinstance(identities, list) and len(identities) == 12, "raw source identity census drift")
    _require(
        source.get("source_members") == 12
        and source.get("source_bytes") == 3_866_949
        and source.get("each_payload_captured_once") is True
        and source.get("parser_consumes_the_same_captured_payload_object") is True,
        "raw source capture contract drift",
    )
    _require(
        source.get("source_identity_ledger_sha256")
        == "ee580671bc9001780fb896e292a58da382c3a1b0b4f14c1adef060f27707c899",
        "raw source identity-ledger digest drift",
    )
    _require(
        hashlib.sha256(witness._canonical_bytes(identities, newline=True)).hexdigest()
        == source["source_identity_ledger_sha256"],
        "raw source identity ledger does not recompute",
    )

    captured: dict[str, CapturedFile] = {}
    for index, embedded in enumerate(identities):
        _require_keys(
            embedded, {"name", "canonical_path", "bytes", "sha256"},
            f"raw TSN PDF identity #{index + 1}",
        )
        item = _bind_embedded_file(embedded, label=f"raw TSN PDF #{index + 1}")
        _require(item.path.suffix.casefold() == ".pdf", f"raw source is not a PDF: {item.path}")
        _require(item.path.name not in captured, f"duplicate raw PDF name: {item.path.name}")
        captured[item.path.name] = item
    non_source_embedded = _require_keys(
        source.get("non_source_role"),
        {"name", "canonical_path", "bytes", "sha256"},
        "raw TSN non-source marker identity",
    )
    non_source = _bind_embedded_file(
        non_source_embedded, label="raw TSN non-source marker"
    )
    _require(non_source.path.name == "_PUT TSN FILES HERE.txt", "raw marker role drift")
    parents = {item.path.parent for item in [*captured.values(), non_source]}
    _require(len(parents) == 1, "raw source members do not share one authoritative root")
    raw_root = next(iter(parents))
    raw_snapshot = _snapshot_tree(raw_root)
    expected_universe = sorted([*captured, non_source.path.name])
    _require(source.get("file_universe") == expected_universe, "raw file universe declaration drift")
    _require(
        [item[0] for item in raw_snapshot.entries] == expected_universe
        and all(item[1] == "file" for item in raw_snapshot.entries),
        "authoritative raw root is not the exact 12-PDF-plus-marker universe",
    )
    expected_snapshot_identities = {
        item.path.name: (item.bytes, item.sha256)
        for item in [*captured.values(), non_source]
    }
    _require(
        {
            str(item[0]): (int(item[2]), str(item[3]))
            for item in raw_snapshot.entries
        } == expected_snapshot_identities,
        "authoritative raw root changed between capture and tree snapshot",
    )
    _require(
        sum(item.bytes for item in captured.values()) == 3_866_949,
        "raw PDF captured byte total drift",
    )
    return {
        **captured,
        non_source.path.name: non_source,
    }, raw_root, raw_snapshot.public()


def _normalized_decision_member(item: Mapping[str, object]) -> dict[str, object]:
    return {
        "name": Path(str(item["canonical_path"])).name,
        "canonical_path": str(Path(str(item["canonical_path"])).resolve(strict=True)),
        "bytes": item.get("bytes", item.get("size")),
        "sha256": item["sha256"],
    }


def _validate_stage6_chain(
    manifest: Mapping[str, object],
    provenance: Mapping[str, object],
    raw_captures: Mapping[str, CapturedFile],
) -> tuple[dict[str, CapturedFile], dict[str, object]]:
    static = _require_keys(
        manifest.get("static_bindings"),
        set(EXPECTED_STATIC_BINDINGS),
        "manifest static_bindings",
    )
    captured: dict[str, CapturedFile] = {}
    for label, expected in EXPECTED_STATIC_BINDINGS.items():
        _require_keys(
            static[label], {"name", "canonical_path", "bytes", "sha256"},
            f"static binding {label} identity",
        )
        captured[label] = _bind_embedded_file(
            static[label], label=f"static binding {label}", expected=expected,
        )

    chain = _require_keys(
        manifest.get("accepted_stage6_chain"),
        {
            "decision", "required_result_flags", "tracked_identities",
            "raw_member_identities_exact", "non_source_role_identity_exact",
        },
        "accepted Stage-6 chain",
    )
    _require(
        provenance.get("accepted_stage6_chain") == chain,
        "manifest/provenance accepted Stage-6 chains disagree",
    )
    required_flags = {
        "normalized_full_conservation": False,
        "projection_exact": False,
        "stage6_family_audit_complete": True,
        "unexplained_projection_residue_count": 0,
    }
    _require(
        chain.get("decision") == "accepted_stage6_family_audit"
        and chain.get("required_result_flags") == required_flags
        and chain.get("raw_member_identities_exact") is True
        and chain.get("non_source_role_identity_exact") is True,
        "accepted Stage-6 chain contract drift",
    )
    tracked = _require_keys(
        chain.get("tracked_identities"), STAGE6_TRACKED_KEYS,
        "accepted Stage-6 tracked identities",
    )
    tracked_captured: dict[str, CapturedFile] = {}
    for label, embedded in tracked.items():
        _require_keys(
            embedded, {"canonical_path", "size", "sha256"},
            f"accepted Stage-6 tracked {label} identity",
        )
        tracked_captured[str(label)] = _bind_embedded_file(
            embedded, label=f"accepted Stage-6 tracked {label}"
        )

    crosswalk = {
        "result": "accepted_stage6_result",
        "oracle": "stage6_oracle",
        "reader": "xlsx_reader",
        "normalized": "accepted_normalized_tsn",
    }
    for tracked_label, static_label in crosswalk.items():
        _require(
            _identity_tuple(tracked_captured[tracked_label])
            == _identity_tuple(captured[static_label]),
            f"Stage-6 tracked/static identity mismatch: {tracked_label}",
        )

    decision = _strict_json_bytes(
        captured["accepted_stage6_decision"].payload,
        "accepted Stage-6 decision", canonical=False,
    )
    _require(isinstance(decision, Mapping), "accepted Stage-6 decision is not an object")
    _require(
        decision.get("decision") == chain["decision"]
        and decision.get("required_result_flags") == required_flags
        and decision.get("tracked_identities") == tracked,
        "detached Stage-6 decision does not bind the embedded accepted chain",
    )
    decision_members = decision.get("raw_member_identities")
    _require(isinstance(decision_members, list) and len(decision_members) == 12, "Stage-6 decision raw census drift")
    observed_pdf_identities = [
        _public_source_identity(item)
        for name, item in raw_captures.items()
        if name.casefold().endswith(".pdf")
    ]
    observed_pdf_identities.sort(key=lambda item: str(item["name"]))
    normalized_decision = [
        _normalized_decision_member(item) for item in decision_members
    ]
    normalized_decision.sort(key=lambda item: str(item["name"]))
    _require(
        normalized_decision == observed_pdf_identities,
        "accepted Stage-6 decision raw identities differ from direct captures",
    )
    non_source_roles = decision.get("non_source_role_identities")
    _require(isinstance(non_source_roles, list) and len(non_source_roles) == 1, "Stage-6 marker decision census drift")
    marker = raw_captures["_PUT TSN FILES HERE.txt"]
    _require(
        _normalized_decision_member(non_source_roles[0])
        == _public_source_identity(marker),
        "accepted Stage-6 marker identity differs from direct capture",
    )

    stage6_result = _strict_json_bytes(
        captured["accepted_stage6_result"].payload,
        "accepted Stage-6 result", canonical=False,
    )
    _require(isinstance(stage6_result, Mapping), "accepted Stage-6 result is not an object")
    for key, expected in required_flags.items():
        _require(stage6_result.get(key) == expected, f"accepted Stage-6 result flag drift: {key}")
    fresh = _require_keys(
        manifest.get("fresh_stage6_digests"), FRESH_STAGE6_KEYS,
        "fresh Stage-6 digest contract",
    )
    _require(fresh.get("all_equal_accepted_stage6") is True, "fresh Stage-6 digest contract absent")
    for key in (
        "raw_source_digests", "raw_source_provenance_digests",
        "document_metadata_digests",
    ):
        _require(
            fresh.get(key) == stage6_result.get(key),
            f"fresh direct-source Stage-6 digest drift: {key}",
        )
    return {**captured, **tracked_captured}, {
        "decision": chain["decision"],
        "required_result_flags": required_flags,
        "tracked_files": len(tracked_captured),
        "static_files": len(captured),
        "raw_member_identities_exact": True,
        "non_source_role_identity_exact": True,
        "fresh_stage6_digests_exact": True,
    }


def _validate_topology(
    result: Mapping[str, object],
    manifest: Mapping[str, object],
    provenance: Mapping[str, object],
) -> dict[str, object]:
    topology = _require_keys(
        manifest.get("bidirectional_equate_topology"),
        {
            "equate_rows", "data_e_rows", "forward_paired", "reverse_paired",
            "forward_unpaired", "reverse_unpaired", "forward_ledger_sha256",
            "reverse_ledger_sha256", "pair_ledgers_exact",
            "forward_unpaired_records", "reverse_unpaired_records", "pairs",
        },
        "bidirectional EQUATES topology",
    )
    _require(
        provenance.get("bidirectional_equate_topology") == topology,
        "manifest/provenance topology ledgers disagree",
    )
    expected_summary = {
        key: topology[key]
        for key in (
            "equate_rows", "data_e_rows", "forward_paired", "reverse_paired",
            "forward_unpaired", "reverse_unpaired", "forward_ledger_sha256",
            "reverse_ledger_sha256", "pair_ledgers_exact",
        )
    }
    _require(
        result.get("bidirectional_equate_topology") == expected_summary,
        "result topology summary disagrees with full topology ledger",
    )
    expected_digest = "2fc7db523404d6d18c0ced944ed24537fa86be240692c6feae15f1223b6eddbd"
    _require(
        topology.get("equate_rows") == 998
        and topology.get("data_e_rows") == 998
        and topology.get("forward_paired") == 998
        and topology.get("reverse_paired") == 998
        and topology.get("forward_unpaired") == 0
        and topology.get("reverse_unpaired") == 0
        and topology.get("forward_ledger_sha256") == expected_digest
        and topology.get("reverse_ledger_sha256") == expected_digest
        and topology.get("pair_ledgers_exact") is True
        and topology.get("forward_unpaired_records") == []
        and topology.get("reverse_unpaired_records") == [],
        "998/998 bidirectional zero-orphan topology drift",
    )
    pairs = topology.get("pairs")
    _require(isinstance(pairs, list) and len(pairs) == 998, "topology pair ledger census drift")
    _require(
        hashlib.sha256(witness._canonical_bytes(pairs, newline=True)).hexdigest()
        == expected_digest,
        "topology pair ledger digest does not recompute",
    )
    pair_keys = {
        "county", "direction", "district", "line", "member", "physical_page",
        "pm", "printed_page", "route", "top",
    }
    seen: set[bytes] = set()
    for index, pair in enumerate(pairs):
        pair = _require_keys(pair, {"annotation", "data_e"}, f"topology pair {index}")
        annotation = _require_keys(pair["annotation"], pair_keys, f"topology annotation {index}")
        data_e = _require_keys(pair["data_e"], pair_keys, f"topology data-E {index}")
        _require(
            all(annotation[key] == data_e[key] for key in ("member", "district", "route", "direction"))
            and str(data_e["pm"]).endswith("E"),
            f"topology ownership/E contract drift at pair {index}",
        )
        encoded = witness._canonical_bytes(pair)
        _require(encoded not in seen, f"duplicate topology pair at index {index}")
        seen.add(encoded)

    mutation = _require_keys(
        manifest.get("reverse_topology_mutation"),
        {
            "label", "rejected", "mutation", "forward_contract_unchanged",
            "reverse_contract_changed", "expected_forward", "observed_forward",
            "expected_reverse", "observed_reverse",
        },
        "reverse topology mutation",
    )
    expected_forward = {
        "equate_rows": 998, "paired": 998, "unpaired": 0,
        "ledger_sha256": expected_digest,
    }
    expected_reverse = {
        "data_e_rows": 998, "paired": 998, "unpaired": 0,
        "ledger_sha256": expected_digest,
    }
    observed_reverse = {
        "data_e_rows": 999, "paired": 998, "unpaired": 1,
        "ledger_sha256": expected_digest,
    }
    planted = _require_keys(
        mutation.get("mutation"), {"source_ref", "before_pm", "after_pm"},
        "reverse topology planted mutation",
    )
    planted_ref = _require_keys(
        planted.get("source_ref"), pair_keys,
        "reverse topology planted source reference",
    )
    before_pm = planted.get("before_pm")
    after_pm = planted.get("after_pm")
    _require(
        mutation.get("label")
        == "unpaired data-E row must be rejected by reverse topology"
        and isinstance(before_pm, str)
        and isinstance(after_pm, str)
        and not before_pm.endswith("E")
        and after_pm == before_pm + "E"
        and planted_ref.get("pm") == after_pm
        and mutation.get("rejected") is True
        and mutation.get("forward_contract_unchanged") is True
        and mutation.get("reverse_contract_changed") is True
        and mutation.get("expected_forward") == expected_forward
        and mutation.get("observed_forward") == expected_forward
        and mutation.get("expected_reverse") == expected_reverse
        and mutation.get("observed_reverse") == observed_reverse
        and result.get("reverse_topology_mutation_rejected") is True,
        "reverse-only orphan-E mutation contract drift",
    )
    return {
        "pairs": 998,
        "forward_paired": 998,
        "reverse_paired": 998,
        "forward_unpaired": 0,
        "reverse_unpaired": 0,
        "ledger_sha256": expected_digest,
        "reverse_only_mutation_rejected": True,
    }


def _validate_provenance_rows(
    provenance: Mapping[str, object],
    source_capture: Mapping[str, object],
    topology: Mapping[str, object],
) -> tuple[list[list[object]], dict[str, object]]:
    schema = provenance.get("schema")
    _require(
        schema == {"sheet": SHEET_NAME, "headers": list(HEADERS)},
        "provenance workbook schema drift",
    )
    rows = provenance.get("rows")
    _require(
        provenance.get("row_count") == 69_804
        and isinstance(rows, list)
        and len(rows) == 69_804,
        "provenance row census drift",
    )
    documents = provenance.get("raw_documents")
    _require(isinstance(documents, list) and len(documents) == 12, "raw document census drift")
    _require(
        all(isinstance(document, Mapping) for document in documents),
        "raw document ledger contains a non-object",
    )
    source_identities = source_capture.get("source_identities")
    _require(isinstance(source_identities, list) and len(source_identities) == 12, "source identity census drift")
    _require(
        all(isinstance(item, Mapping) and isinstance(item.get("name"), str)
            for item in source_identities),
        "source identity ledger contains an invalid member",
    )
    member_by_document = {
        index: str(item["name"])
        for index, item in enumerate(source_identities, 1)
    }
    _require(
        [document.get("member") for document in documents]
        == [member_by_document[index] for index in range(1, 13)],
        "raw document/source member order drift",
    )

    workbook_rows: list[list[object]] = []
    global_ordinals: set[int] = set()
    member_ordinals: set[tuple[int, int]] = set()
    member_ordinal_values: dict[int, set[int]] = {
        document: set() for document in range(1, 13)
    }
    kinds: Counter[str] = Counter()
    pointer_tokens: Counter[object] = Counter()
    route_counts: Counter[str] = Counter()
    blank_county = 0
    equate_refs: set[bytes] = set()
    data_e_refs: set[bytes] = set()
    expected_row_keys = {
        "workbook_row", "source_ordinals", "source_ref", "source_context",
        "workbook_values",
    }
    source_ordinal_keys = {
        "document_ordinal", "member_record_ordinal", "global_parse_ordinal",
    }
    source_ref_keys = {
        "member", "physical_page", "printed_page", "line", "top",
    }
    source_context_keys = {
        "district", "route", "direction", "kind", "raw_text",
    }
    topology_ref_keys = {
        "member", "district", "route", "direction", "county", "pm",
        "physical_page", "printed_page", "line", "top",
    }

    for index, raw_row in enumerate(rows):
        row = _require_keys(raw_row, expected_row_keys, f"provenance row {index + 2}")
        _require(row.get("workbook_row") == index + 2, f"provenance workbook-row sequence drift at {index + 2}")
        ordinals = _require_keys(
            row.get("source_ordinals"), source_ordinal_keys,
            f"provenance ordinals {index + 2}",
        )
        source_ref = _require_keys(
            row.get("source_ref"), source_ref_keys,
            f"provenance source_ref {index + 2}",
        )
        context = _require_keys(
            row.get("source_context"), source_context_keys,
            f"provenance source_context {index + 2}",
        )
        values = row.get("workbook_values")
        _require(
            isinstance(values, list)
            and len(values) == 8
            and all(value is None or isinstance(value, str) for value in values),
            f"provenance workbook values are not eight typed string/blank cells at {index + 2}",
        )
        document_ordinal = ordinals.get("document_ordinal")
        member_ordinal = ordinals.get("member_record_ordinal")
        global_ordinal = ordinals.get("global_parse_ordinal")
        _require(
            type(document_ordinal) is int and 1 <= document_ordinal <= 12
            and type(member_ordinal) is int and member_ordinal > 0
            and type(global_ordinal) is int and 1 <= global_ordinal <= 69_804,
            f"invalid source ordinal at workbook row {index + 2}",
        )
        _require(
            source_ref.get("member") == member_by_document[document_ordinal],
            f"document ordinal/member disagreement at workbook row {index + 2}",
        )
        _require(global_ordinal not in global_ordinals, f"duplicate global parse ordinal: {global_ordinal}")
        ordinal_pair = (document_ordinal, member_ordinal)
        _require(ordinal_pair not in member_ordinals, f"duplicate member source ordinal: {ordinal_pair}")
        global_ordinals.add(global_ordinal)
        member_ordinals.add(ordinal_pair)
        member_ordinal_values[document_ordinal].add(member_ordinal)

        kind = context.get("kind")
        _require(kind in {"data", "equate"}, f"unknown provenance kind at row {index + 2}")
        _require(
            values[0] == context.get("route"),
            f"workbook/provenance Route disagreement at row {index + 2}",
        )
        kinds[str(kind)] += 1
        route_counts[str(values[0])] += 1
        pointer_tokens[values[6]] += 1
        if values[1] is None:
            blank_county += 1
            _require(kind == "equate", f"blank County is not EQUATES at row {index + 2}")

        topology_ref = {
            "member": source_ref["member"],
            "district": context["district"],
            "route": context["route"],
            "direction": context["direction"],
            "county": values[1],
            "pm": values[2],
            "physical_page": source_ref["physical_page"],
            "printed_page": source_ref["printed_page"],
            "line": source_ref["line"],
            "top": source_ref["top"],
        }
        _require(set(topology_ref) == topology_ref_keys, "internal topology-ref shape error")
        encoded_ref = witness._canonical_bytes(topology_ref)
        if kind == "equate":
            equate_refs.add(encoded_ref)
        if kind == "data" and str(values[2]).endswith("E"):
            data_e_refs.add(encoded_ref)
        workbook_rows.append(values)

    _require(global_ordinals == set(range(1, 69_805)), "global source ordinals are not exhaustive")
    _require(
        all(
            values == set(range(1, len(values) + 1))
            for values in member_ordinal_values.values()
        ),
        "member source ordinals are not independently contiguous",
    )
    _require(kinds == Counter({"data": 68_806, "equate": 998}), "provenance kind census drift")
    _require(blank_county == 46, "provenance blank-County EQUATES census drift")
    _require(
        pointer_tokens["*P*"] == 283
        and pointer_tokens["-------->"] == 282,
        "provenance pointer-token census drift",
    )
    ordered_sha = _ordered_rows_sha256(workbook_rows)
    _require(ordered_sha == EXPECTED_ORDERED_ROWS_SHA256, "provenance ordered-row digest drift")
    pairs = topology["pairs"]
    pair_annotations = {
        witness._canonical_bytes(pair["annotation"]) for pair in pairs
    }
    pair_data_e = {
        witness._canonical_bytes(pair["data_e"]) for pair in pairs
    }
    _require(
        pair_annotations == equate_refs
        and pair_data_e == data_e_refs
        and len(equate_refs) == len(data_e_refs) == 998,
        "topology pair ledger is not exhaustive over provenance EQUATES/data-E rows",
    )
    return workbook_rows, {
        "rows": len(workbook_rows),
        "data_records": kinds["data"],
        "equate_records": kinds["equate"],
        "blank_county_equates": blank_county,
        "pointer_P": pointer_tokens["*P*"],
        "pointer_arrow": pointer_tokens["-------->"],
        "pointer_total": pointer_tokens["*P*"] + pointer_tokens["-------->"],
        "ordered_rows_sha256": ordered_sha,
        "route_counts": dict(sorted(route_counts.items())),
        "global_source_ordinals_exact": True,
        "member_source_ordinals_unique": True,
        "topology_refs_exhaustive": True,
    }


def _validate_canonical_package(
    workbook: CapturedFile,
    manifest: Mapping[str, object],
) -> dict[str, object]:
    package_contract = _require_keys(
        manifest.get("canonical_xlsx_package"),
        {
            "package_members", "package_member_names_sorted",
            "fixed_zip_timestamps", "empty_zip_member_extra_and_comments",
            "core_timestamp_contract", "member_ledger_sha256", "members",
        },
        "canonical XLSX package contract",
    )
    try:
        with ZipFile(io.BytesIO(workbook.payload), "r") as package:
            infos = package.infolist()
            names = package.namelist()
            _require(package.testzip() is None, "direct twin XLSX CRC failure")
            _require(package.comment == b"", "direct twin XLSX archive comment drift")
            _require(len(names) == len(set(names)), "direct twin XLSX has duplicate members")
            required_parts = {
                "[Content_Types].xml", "_rels/.rels", "docProps/core.xml",
                "xl/workbook.xml", "xl/_rels/workbook.xml.rels",
                "xl/worksheets/sheet1.xml", "xl/styles.xml",
            }
            _require(required_parts <= set(names), "direct twin XLSX required-part gap")
            member_payloads = {name: package.read(name) for name in names}
            members = [
                {
                    "name": name,
                    "bytes": len(member_payloads[name]),
                    "sha256": hashlib.sha256(member_payloads[name]).hexdigest(),
                }
                for name in sorted(names)
            ]
            fixed_timestamps = all(info.date_time == (1980, 1, 1, 0, 0, 0) for info in infos)
            empty_metadata = all(not info.extra and not info.comment for info in infos)
    except (OSError, ValueError) as exc:
        raise DirectRawWitnessError("direct twin workbook is not a valid XLSX package") from exc
    try:
        core_root = ElementTree.fromstring(member_payloads["docProps/core.xml"])
    except ElementTree.ParseError as exc:
        raise DirectRawWitnessError("direct twin core properties XML is invalid") from exc
    dcterms = "http://purl.org/dc/terms/"
    created = core_root.findall(f".//{{{dcterms}}}created")
    modified = core_root.findall(f".//{{{dcterms}}}modified")
    _require(len(created) == len(modified) == 1, "direct twin core timestamp node census drift")
    core_contract = {
        "created_nodes": 1,
        "modified_nodes": 1,
        "created": created[0].text,
        "modified": modified[0].text,
        "exact_fixed": (
            created[0].text == "2000-01-01T00:00:00Z"
            and modified[0].text == "2000-01-01T00:00:00Z"
        ),
    }
    recomputed = {
        "package_members": len(members),
        "package_member_names_sorted": names == sorted(names),
        "fixed_zip_timestamps": fixed_timestamps,
        "empty_zip_member_extra_and_comments": empty_metadata,
        "core_timestamp_contract": core_contract,
        "member_ledger_sha256": hashlib.sha256(
            witness._canonical_bytes(members, newline=True)
        ).hexdigest(),
        "members": members,
    }
    _require(
        all(recomputed[key] == package_contract.get(key) for key in recomputed),
        "canonical XLSX package ledger/metadata drift",
    )
    _require(
        recomputed["package_member_names_sorted"] is True
        and recomputed["fixed_zip_timestamps"] is True
        and recomputed["empty_zip_member_extra_and_comments"] is True,
        "direct twin XLSX is not byte-canonical",
    )
    _require(core_contract["exact_fixed"] is True, "direct twin core timestamps are volatile")
    return {
        "members": recomputed["package_members"],
        "member_ledger_sha256": recomputed["member_ledger_sha256"],
        "fixed_zip_timestamps": True,
        "empty_zip_metadata": True,
        "core_timestamp_contract": core_contract,
    }


def _validate_builder_determinism(
    manifest: Mapping[str, object],
    workbook: CapturedFile,
    package_summary: Mapping[str, object],
) -> dict[str, object]:
    gate = _require_keys(
        manifest.get("output_root_gate"),
        {
            "valid_private_child_accepted", "raw_root_child_rejected",
            "static_artifact_parent_child_rejected",
            "lexical_parent_alias_rejected",
            "multi_level_reparse_component_rejected", "reparse_control",
        },
        "direct-twin builder output-root controls",
    )
    reparse_control = _require_keys(
        gate.get("reparse_control"),
        {
            "kind", "alias_is_multiple_components_above_output",
            "rejected_before_resolve",
        },
        "direct-twin builder reparse control",
    )
    _require(
        all(gate[key] is True for key in (
            "valid_private_child_accepted", "raw_root_child_rejected",
            "static_artifact_parent_child_rejected",
            "lexical_parent_alias_rejected",
            "multi_level_reparse_component_rejected",
        ))
        and reparse_control == {
            "kind": "directory_symlink_reparse_point",
            "alias_is_multiple_components_above_output": True,
            "rejected_before_resolve": True,
        },
        "direct-twin builder output/reparse controls drift",
    )
    workbook_identity = {"bytes": workbook.bytes, "sha256": workbook.sha256}
    delayed = _require_keys(
        manifest.get("delayed_in_memory_determinism"),
        {
            "delay_milliseconds", "first", "second",
            "package_contract_exact", "byte_identical",
        },
        "delayed in-memory determinism contract",
    )
    _require(
        delayed.get("delay_milliseconds") == 1_200
        and delayed.get("first") == workbook_identity
        and delayed.get("second") == workbook_identity
        and delayed.get("package_contract_exact") is True
        and delayed.get("byte_identical") is True,
        "delayed in-memory workbook determinism drift",
    )
    mutation = _require_keys(
        manifest.get("core_timestamp_mutation"),
        {
            "label", "planted_modified", "expected_contract", "mutated_contract",
            "rejected_by_exact_contract", "recanonicalized_byte_identical",
            "mutated_artifact", "recovered_artifact",
            "recovered_member_ledger_sha256",
        },
        "core timestamp mutation contract",
    )
    expected_core = package_summary["core_timestamp_contract"]
    mutated_core = {
        "created_nodes": 1,
        "modified_nodes": 1,
        "created": "2000-01-01T00:00:00Z",
        "modified": "2001-02-03T04:05:06Z",
        "exact_fixed": False,
    }
    _require(
        mutation.get("label") == "volatile docProps/core.xml modified timestamp"
        and mutation.get("planted_modified") == "2001-02-03T04:05:06Z"
        and mutation.get("expected_contract") == expected_core
        and mutation.get("mutated_contract") == mutated_core
        and mutation.get("rejected_by_exact_contract") is True
        and mutation.get("recanonicalized_byte_identical") is True
        and mutation.get("recovered_artifact") == workbook_identity
        and mutation.get("recovered_member_ledger_sha256")
        == package_summary["member_ledger_sha256"]
        and isinstance(mutation.get("mutated_artifact"), Mapping)
        and mutation["mutated_artifact"].get("sha256") != workbook.sha256,
        "core-timestamp mutation/recanonicalization contract drift",
    )
    return {
        "output_root_controls_exact": True,
        "multi_level_reparse_control_exact": True,
        "delayed_rebuild_byte_identical": True,
        "core_timestamp_exact_fixed": True,
        "core_timestamp_mutation_rejected_and_recovered": True,
    }


def _validate_workbook(
    workbook: CapturedFile,
    expected_rows: Sequence[Sequence[object]],
    manifest: Mapping[str, object],
) -> dict[str, object]:
    _require(
        manifest.get("workbook_schema")
        == {"sheet": SHEET_NAME, "headers": list(HEADERS)},
        "manifest workbook schema drift",
    )
    package = _validate_canonical_package(workbook, manifest)
    try:
        opened = load_workbook(
            io.BytesIO(workbook.payload), read_only=True, data_only=False,
        )
    except Exception as exc:
        raise DirectRawWitnessError("cannot reopen captured direct twin workbook") from exc
    try:
        _require(opened.sheetnames == [SHEET_NAME], "direct twin workbook sheet universe drift")
        sheet = opened[SHEET_NAME]
        physical_rows = 0
        data_rows = 0
        formulas = errors = blank_rows = extra_cells = padded = 0
        maximum_width = 0
        digest = hashlib.sha256()
        render_surface: list[list[object]] = []
        for physical_row, cells in enumerate(sheet.iter_rows(), 1):
            physical_rows += 1
            maximum_width = max(maximum_width, len(cells))
            extra_cells += max(0, len(cells) - 8)
            values = tuple(cell.value for cell in cells[:8])
            formulas += sum(cell.data_type == "f" for cell in cells)
            errors += sum(cell.data_type == "e" for cell in cells)
            if len(values) < 8:
                padded += 8 - len(values)
                values += (None,) * (8 - len(values))
            if not any(value is not None for value in values):
                blank_rows += 1
            if physical_row == 1:
                _require(values == HEADERS, "direct twin workbook header drift")
            else:
                data_rows += 1
                _require(
                    data_rows <= len(expected_rows)
                    and values == tuple(expected_rows[data_rows - 1]),
                    f"workbook/provenance typed-row disagreement at workbook row {physical_row}",
                )
                digest.update(json.dumps(
                    list(values), ensure_ascii=False, separators=(",", ":"),
                ).encode("utf-8"))
                digest.update(b"\n")
            if physical_row <= 12:
                render_surface.append(list(values))
        ordered_sha = digest.hexdigest()
    finally:
        opened.close()

    try:
        styled = load_workbook(
            io.BytesIO(workbook.payload), read_only=False, data_only=False,
        )
    except Exception as exc:
        raise DirectRawWitnessError("cannot reopen direct twin workbook styling") from exc
    try:
        sheet = styled[SHEET_NAME]
        header = next(sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8))
        header_style_exact = all(
            cell.font.bold
            and cell.font.color is not None
            and cell.fill.fill_type == "solid"
            for cell in header
        )
        freeze_panes = sheet.freeze_panes
        auto_filter = sheet.auto_filter.ref
        description_width = sheet.column_dimensions["H"].width
        gridlines_hidden = sheet.sheet_view.showGridLines is False
    finally:
        styled.close()

    observed = {
        "zip_test_passed": True,
        "required_parts_present": True,
        "duplicate_package_members": 0,
        "sheet_names_exact": True,
        "sheet_name": SHEET_NAME,
        "streamed_physical_rows": physical_rows,
        "streamed_data_rows": data_rows,
        "streamed_logical_columns": 8,
        "maximum_streamed_physical_width": maximum_width,
        "padded_omitted_trailing_blank_cells": padded,
        "extra_physical_cells": extra_cells,
        "blank_physical_rows": blank_rows,
        "formulas": formulas,
        "errors": errors,
        "typed_rows_exact": True,
        "ordered_rows_sha256": ordered_sha,
        "freeze_panes": freeze_panes,
        "auto_filter": auto_filter,
        "render_surface_probe": {
            "range": "A1:H12",
            "values_sha256": hashlib.sha256(
                witness._canonical_bytes(render_surface, newline=True)
            ).hexdigest(),
            "header_style_exact": header_style_exact,
            "freeze_panes": freeze_panes,
            "auto_filter": auto_filter,
            "description_column_width": description_width,
            "gridlines_hidden": gridlines_hidden,
        },
    }
    declared = manifest.get("workbook_reopen")
    _require(observed == declared, "independent canonical workbook reopen disagrees with manifest")
    _require(
        physical_rows == 69_805
        and data_rows == 69_804
        and maximum_width == 8
        and padded == 13_468
        and not formulas and not errors and not blank_rows and not extra_cells
        and ordered_sha == EXPECTED_ORDERED_ROWS_SHA256
        and freeze_panes == "A2"
        and auto_filter == "A1:H69805"
        and header_style_exact
        and description_width == 72.0
        and gridlines_hidden,
        "direct twin workbook extent/render/schema contract drift",
    )
    return {
        **package,
        "sheet": SHEET_NAME,
        "physical_rows": physical_rows,
        "data_rows": data_rows,
        "logical_columns": 8,
        "padded_trailing_blanks": padded,
        "formulas": formulas,
        "errors": errors,
        "ordered_rows_sha256": ordered_sha,
        "typed_rows_equal_provenance": True,
    }


def _coalesce_protected_roots(roots: Iterable[Path]) -> tuple[Path, ...]:
    canonical = sorted(
        {path.resolve(strict=True) for path in roots},
        key=lambda path: (len(path.parts), str(path).casefold()),
    )
    selected: list[Path] = []
    for candidate in canonical:
        if candidate == VISUAL_ROOT.resolve(strict=True):
            continue
        if any(existing == candidate or existing in candidate.parents for existing in selected):
            continue
        selected.append(candidate)
    return tuple(selected)


def _assert_product_not_loaded() -> None:
    loaded = sorted(
        name for name in sys.modules
        if name in PRODUCT_MODULES
        or any(name.endswith(f".{product}") for product in PRODUCT_MODULES)
    )
    _require(not loaded, f"product code loaded before direct-source validation: {loaded!r}")


def _audit_code_reported_paths() -> dict[str, Path]:
    return {
        "direct_product_runner": RUNNER_LEXICAL_PATH,
        "shared_product_witness": Path(str(witness.__file__)),
        "raw_chunk_witness": Path(str(dev.__file__)),
        "direct_twin_builder": DIRECT_BUILDER,
    }


def _capture_audit_code(
    *, reported_paths: Mapping[str, Path] | None = None,
) -> dict[str, CapturedFile]:
    supplied = _audit_code_reported_paths() if reported_paths is None else reported_paths
    _require(
        set(supplied) == set(AUDIT_CODE_LOGICAL_PATHS),
        "audit-code reported role universe drift",
    )
    logical_paths = {
        label: _lexical_absolute(REPO_ROOT / Path(logical_path))
        for label, logical_path in AUDIT_CODE_LOGICAL_PATHS.items()
    }
    # Capture the declared logical path first.  _capture_file performs component
    # lstat/reparse rejection before its eventual canonical-path lookup.
    captured = {
        label: _capture_file(path) for label, path in logical_paths.items()
    }
    for label in sorted(supplied):
        reported = _lexical_abspath(Path(supplied[label]))
        reported_capture = _capture_file(reported)
        logical_capture = captured[label]
        _require(
            _frozen_identity_tuple(reported_capture)
            == _frozen_identity_tuple(logical_capture),
            f"audit-code physical/logical path attribution drift: {label}",
        )
    _require(
        (captured["direct_twin_builder"].bytes,
         captured["direct_twin_builder"].sha256)
        == PINNED_BUILDER_IDENTITY,
        "audit-code set contains an unpinned direct-twin builder",
    )
    _require(
        len({item.path for item in captured.values()}) == len(captured),
        "audit-code roles alias one physical path",
    )
    return captured


def _audit_code_manifest(
    captured: Mapping[str, CapturedFile],
) -> dict[str, object]:
    _require(
        set(captured) == set(AUDIT_CODE_LOGICAL_PATHS),
        "audit-code role universe drift",
    )
    members = [
        {
            "role": label,
            "logical_path": AUDIT_CODE_LOGICAL_PATHS[label],
            "bytes": captured[label].bytes,
            "sha256": captured[label].sha256,
        }
        for label in sorted(captured)
    ]
    return {
        "schema": "phase8-direct-raw-audit-code/v1",
        "roles": len(members),
        "canonical_members_sha256": hashlib.sha256(
            witness._canonical_bytes(members)
        ).hexdigest(),
        "members": members,
    }


def _revalidate_audit_code(
    baseline: Mapping[str, CapturedFile],
) -> dict[str, object]:
    observed = {label: _capture_file(item.path) for label, item in baseline.items()}
    _require(
        {
            label: _frozen_identity_tuple(item)
            for label, item in observed.items()
        } == {
            label: _frozen_identity_tuple(item)
            for label, item in baseline.items()
        },
        "audit-code identity changed during direct product witness",
    )
    return _audit_code_manifest(observed)


def _run_audit_code_mutation_probe(
    source: Mapping[str, CapturedFile],
) -> dict[str, object]:
    with tempfile.TemporaryDirectory(
        prefix="phase8-direct-audit-code-control-", dir=VISUAL_ROOT,
    ) as temporary:
        root = Path(temporary).resolve(strict=True)
        baseline: dict[str, CapturedFile] = {}
        for label, captured in source.items():
            copy = root / f"{label}.py"
            copy.write_bytes(captured.payload)
            baseline[label] = _capture_file(copy)
        copied_baseline_manifest = _revalidate_audit_code(baseline)
        _require(
            copied_baseline_manifest["roles"] == 4,
            "unchanged copied audit-code baseline did not pass",
        )
        rejected_roles: list[str] = []
        for label in sorted(baseline):
            selected = root / f"{label}.py"
            selected.write_bytes(
                source[label].payload
                + f"\n# planted code drift: {label}\n".encode("ascii")
            )
            rejected = False
            try:
                _revalidate_audit_code(baseline)
            except DirectRawWitnessError:
                rejected = True
            _require(rejected, f"audit-code role byte mutation escaped: {label}")
            rejected_roles.append(label)
            selected.write_bytes(source[label].payload)
            restored = _capture_file(selected)
            _require(
                (restored.bytes, restored.sha256)
                == (baseline[label].bytes, baseline[label].sha256),
                f"audit-code mutation control did not restore bytes: {label}",
            )
            baseline[label] = restored
        rejected_redirections: list[str] = []
        production_reported = _audit_code_reported_paths()
        for label in sorted(source):
            redirect = root / f"{label}.reported-path-redirect.py"
            try:
                os.symlink(source[label].path, redirect, target_is_directory=False)
            except (OSError, NotImplementedError) as exc:
                raise DirectRawWitnessError(
                    f"audit-code reported-path reparse mutation could not execute: {label}"
                ) from exc
            redirected = dict(production_reported)
            redirected[label] = redirect
            rejected = False
            try:
                _capture_audit_code(reported_paths=redirected)
            except DirectRawWitnessError:
                rejected = True
            finally:
                try:
                    redirect.unlink()
                except OSError as exc:
                    raise DirectRawWitnessError(
                        f"audit-code reported-path reparse residue remained: {label}"
                    ) from exc
            _require(
                rejected,
                f"audit-code reported-path reparse redirection escaped: {label}",
            )
            rejected_redirections.append(label)
        _require(
            all(
                _frozen_identity_tuple(_capture_file(item.path))
                == _frozen_identity_tuple(item)
                for label, item in source.items()
            ),
            "audit-code mutation probe changed a real helper",
        )
        return {
            "roles_copied": sorted(source),
            "unchanged_copied_baseline_passed": True,
            "mutated_roles": rejected_roles,
            "every_copied_role_byte_mutation_rejected":
                set(rejected_roles) == set(source),
            "reported_role_reparse_redirections_rejected": rejected_redirections,
            "every_reported_role_reparse_redirection_rejected":
                set(rejected_redirections) == set(source),
            "physical_logical_paths_exact": True,
            "real_audit_code_unchanged": True,
        }


def _validate_direct_twin(twin_root: Path) -> Preflight:
    _assert_product_not_loaded()
    _require(
        PINNED_BUILDER_IDENTITY is not None,
        "final direct-twin builder identity has not been pinned after review",
    )
    directory_token_mutations = _run_directory_token_mutation_probes()
    canonical_twin, twin_files, twin_snapshot = _capture_direct_twin_root(twin_root)
    result = _strict_json_bytes(
        twin_files[TWIN_RESULT_NAME].payload, TWIN_RESULT_NAME, canonical=True,
    )
    manifest = _strict_json_bytes(
        twin_files[MANIFEST_NAME].payload, MANIFEST_NAME, canonical=True,
    )
    provenance = _strict_json_bytes(
        twin_files[PROVENANCE_NAME].payload, PROVENANCE_NAME, canonical=True,
    )
    result = _require_keys(result, DIRECT_RESULT_KEYS, "direct twin result")
    manifest = _require_keys(manifest, DIRECT_MANIFEST_KEYS, "direct twin manifest")
    provenance = _require_keys(
        provenance, DIRECT_PROVENANCE_KEYS, "direct twin provenance",
    )
    schema_mutations = _run_v1_schema_mutation_probes(
        result, manifest, provenance,
    )
    _validate_nonacceptance_contract(result, label="direct twin result", terminal_result=True)
    _validate_nonacceptance_contract(
        manifest, label="direct twin manifest", reason_required=False,
    )
    _validate_nonacceptance_contract(provenance, label="direct twin provenance")
    _require(
        result.get("audit") == "Highway Sequence direct-source raw-TSN audit twin builder",
        "direct twin result audit role drift",
    )
    _require(
        manifest.get("audit") == "Highway Sequence direct-source raw-TSN audit twin manifest",
        "direct twin manifest audit role drift",
    )
    _require(
        provenance.get("audit") == "Highway Sequence direct-source raw-TSN audit twin provenance",
        "direct twin provenance audit role drift",
    )
    _require(
        result.get("artifact_universe") == list(TWIN_OUTPUT_NAMES)
        and manifest.get("output_artifact_names") == list(TWIN_OUTPUT_NAMES),
        "direct twin declared artifact universe drift",
    )
    _require(
        result.get("output_root_embedded") is False
        and manifest.get("output_root_embedded") is False,
        "direct twin embeds its volatile output root",
    )
    encoded_root = str(canonical_twin).casefold().encode("utf-8")
    _require(
        all(encoded_root not in captured.payload.lower() for captured in twin_files.values()),
        "direct twin payload embeds its volatile output root path",
    )

    independence = _require_keys(
        manifest.get("independence"),
        {
            "development_row_cache_read", "development_twin_read",
            "product_code_imported", "source_parser",
        },
        "direct twin independence contract",
    )
    _require(
        independence.get("development_row_cache_read") is False
        and independence.get("development_twin_read") is False
        and independence.get("product_code_imported") is False
        and independence.get("source_parser")
        == "exact-bound accepted audit-owned Stage-6 parser",
        "direct twin independence contract drift",
    )
    runtime = _require_keys(
        manifest.get("runtime"), {"python", "openpyxl", "pdfplumber"},
        "direct twin runtime",
    )
    _require(
        isinstance(runtime.get("python"), str)
        and runtime.get("openpyxl") == "3.1.5"
        and runtime.get("pdfplumber") == "0.11.9",
        "direct twin audit-runtime identity drift",
    )
    invariants = _require_keys(
        result.get("invariants"), DIRECT_INVARIANT_KEYS,
        "direct twin terminal invariants",
    )
    _require(
        all(value is True for value in invariants.values()),
        "direct twin result invariants are not all terminal true",
    )

    artifacts = _require_keys(
        result.get("artifacts"), {"workbook", "provenance", "manifest"},
        "direct twin result artifact identities",
    )
    generated = _require_keys(
        manifest.get("generated"), {"workbook", "provenance"},
        "direct twin generated identities",
    )
    artifact_map = {
        "workbook": twin_files[WORKBOOK_NAME],
        "provenance": twin_files[PROVENANCE_NAME],
        "manifest": twin_files[MANIFEST_NAME],
    }
    for label, captured in artifact_map.items():
        _require_keys(
            artifacts[label], {"bytes", "sha256"},
            f"direct twin result {label} identity",
        )
        _require(
            _artifact_identity_matches(artifacts[label], captured),
            f"direct twin result {label} identity drift",
        )
        if label != "manifest":
            _require_keys(
                generated[label], {"bytes", "sha256"},
                f"direct twin manifest {label} identity",
            )
            _require(
                _artifact_identity_matches(generated[label], captured),
                f"direct twin manifest {label} identity drift",
            )

    builder = _capture_file(DIRECT_BUILDER)
    _require(
        (builder.bytes, builder.sha256) == PINNED_BUILDER_IDENTITY,
        "current direct-twin builder differs from the final pinned identity",
    )
    _require_keys(
        manifest.get("builder_identity"),
        {"name", "canonical_path", "bytes", "sha256"},
        "direct twin builder identity",
    )
    _require(
        _embedded_identity_matches(manifest.get("builder_identity"), builder),
        "direct twin was not produced by the exact current builder bytes",
    )

    counts = _require_keys(
        result.get("counts"), DIRECT_COUNT_KEYS, "direct twin counts",
    )
    _require(manifest.get("counts") == counts, "manifest/result count ledgers disagree")
    for key, expected in EXPECTED_COUNTS.items():
        _require(counts.get(key) == expected, f"direct twin count drift: {key}")
    route_counts = counts.get("route_counts")
    _require(
        isinstance(route_counts, Mapping)
        and all(isinstance(key, str) and type(value) is int and value > 0
                for key, value in route_counts.items())
        and sum(route_counts.values()) == 69_804,
        "direct twin route census drift",
    )
    _require(
        result.get("ordered_rows_sha256") == EXPECTED_ORDERED_ROWS_SHA256,
        "direct twin result ordered-row digest drift",
    )

    source_capture = manifest.get("source_capture")
    _require(isinstance(source_capture, Mapping), "direct twin source capture absent")
    raw_captures, raw_root, raw_snapshot_summary = _validate_source_capture(
        manifest, provenance,
    )
    stage6_captures, stage6_summary = _validate_stage6_chain(
        manifest, provenance, raw_captures,
    )
    topology_summary = _validate_topology(result, manifest, provenance)
    topology = manifest["bidirectional_equate_topology"]
    workbook_rows, provenance_summary = _validate_provenance_rows(
        provenance, source_capture, topology,
    )
    _require(
        provenance_summary["route_counts"] == route_counts,
        "provenance/result route-count ledgers disagree",
    )
    workbook_summary = _validate_workbook(
        twin_files[WORKBOOK_NAME], workbook_rows, manifest,
    )
    builder_determinism = _validate_builder_determinism(
        manifest, twin_files[WORKBOOK_NAME], workbook_summary,
    )

    tsmis_captures: dict[str, CapturedFile] = {}
    for label, (path, expected_bytes, expected_sha) in EXPECTED_TSMIS.items():
        captured = _capture_file(path)
        _require(
            (captured.bytes, captured.sha256) == (expected_bytes, expected_sha),
            f"exact TSMIS {label} input identity drift",
        )
        tsmis_captures[label] = captured
    source_snapshot = _snapshot_tree(SOURCE_ROOT)
    source_public = source_snapshot.public()
    _require(
        source_public["files"] == 255
        and source_public["directories"] == 1
        and source_public["bytes"] == 8_572_971
        and source_public["canonical_members_sha256"]
        == EXPECTED_SOURCE_TREE_LEDGER_SHA256,
        "current TSMIS product-source tree universe drift",
    )
    source_entry_identities = {
        str(item[0]): (int(item[2]), str(item[3]))
        for item in source_snapshot.entries if item[1] == "file"
    }
    for captured in tsmis_captures.values():
        relative = captured.path.relative_to(SOURCE_ROOT.resolve(strict=True)).as_posix()
        _require(
            source_entry_identities.get(relative)
            == (captured.bytes, captured.sha256),
            f"TSMIS input changed between capture and source-tree snapshot: {relative}",
        )

    all_captures: list[CapturedFile] = [
        *twin_files.values(), builder, *tsmis_captures.values(),
        *raw_captures.values(), *stage6_captures.values(),
    ]
    by_path: dict[Path, CapturedFile] = {}
    for captured in all_captures:
        prior = by_path.get(captured.path)
        if prior is not None:
            _require(
                _frozen_identity_tuple(prior) == _frozen_identity_tuple(captured),
                f"same bound path produced inconsistent captures: {captured.path}",
            )
        by_path[captured.path] = captured

    visual = VISUAL_ROOT.resolve(strict=True)
    artifact_root_candidates: set[Path] = {
        canonical_twin, SOURCE_ROOT.resolve(strict=True), raw_root,
    }
    for path in by_path:
        if path == canonical_twin or canonical_twin in path.parents:
            continue
        if path == SOURCE_ROOT or SOURCE_ROOT.resolve(strict=True) in path.parents:
            continue
        try:
            path.relative_to(visual)
        except ValueError:
            continue
        artifact_root_candidates.add(path.parent)
    protected_roots = _coalesce_protected_roots(artifact_root_candidates)
    snapshots_by_root: dict[Path, TreeSnapshot] = {
        twin_snapshot.root: twin_snapshot,
        source_snapshot.root: source_snapshot,
    }
    raw_snapshot = _snapshot_tree(raw_root)
    snapshots_by_root[raw_snapshot.root] = raw_snapshot
    for root in protected_roots:
        if root not in snapshots_by_root:
            snapshots_by_root[root] = _snapshot_tree(root)
    snapshots = tuple(
        snapshots_by_root[root]
        for root in sorted(snapshots_by_root, key=lambda path: str(path).casefold())
    )
    for captured in by_path.values():
        containing = [
            snapshot for snapshot in snapshots
            if captured.path == snapshot.root or snapshot.root in captured.path.parents
        ]
        if not containing:
            continue
        snapshot = max(containing, key=lambda item: len(item.root.parts))
        _require(captured.path != snapshot.root, f"bound file aliases input-tree root: {captured.path}")
        relative = captured.path.relative_to(snapshot.root).as_posix()
        matches = [item for item in snapshot.entries if item[0] == relative]
        _require(
            len(matches) == 1
            and matches[0][1] == "file"
            and (matches[0][2], matches[0][3], matches[0][4])
            == (captured.bytes, captured.sha256, captured.token),
            f"bound file/tree snapshot identity disagreement: {captured.path}",
        )
    bound_identities = tuple(sorted(
        (_frozen_identity_tuple(captured) for captured in by_path.values()),
        key=lambda item: item[0].casefold(),
    ))
    summary = {
        "validated_before_product_import": True,
        "schema_version": 1,
        "status": result["status"],
        "artifact_status": result["artifact_status"],
        "acceptance_eligible": False,
        "stage8_family_accepted": False,
        "twin_root": str(canonical_twin),
        "artifact_universe": list(TWIN_OUTPUT_NAMES),
        "artifacts": {
            name: twin_files[name].identity() for name in TWIN_OUTPUT_NAMES
        },
        "builder": builder.identity(),
        "counts": dict(counts),
        "provenance": provenance_summary,
        "workbook": workbook_summary,
        "builder_determinism": builder_determinism,
        "directory_token_mutations": directory_token_mutations,
        "v1_schema_mutations": schema_mutations,
        "topology": topology_summary,
        "stage6_chain": stage6_summary,
        "raw_source_tree": raw_snapshot_summary,
        "tsmis_inputs": {
            label: captured.identity() for label, captured in tsmis_captures.items()
        },
        "frozen_input_trees": [snapshot.public() for snapshot in snapshots],
        "bound_files": len(by_path),
        "direct_twin_current_builder_exact": True,
        "not_family_acceptance": True,
    }
    _assert_product_not_loaded()
    return Preflight(
        summary=summary,
        bound_files=tuple(sorted(by_path, key=lambda path: str(path).casefold())),
        protected_roots=protected_roots,
        tree_snapshots=snapshots,
        bound_identities=bound_identities,
        raw_tsn_workbook=twin_files[WORKBOOK_NAME].path,
    )


def _revalidate_frozen_inputs(preflight: Preflight, *, phase: str) -> dict[str, object]:
    rebound = tuple(sorted(
        (_frozen_identity_tuple(_capture_file(Path(item[0])))
         for item in preflight.bound_identities),
        key=lambda item: item[0].casefold(),
    ))
    _require(rebound == preflight.bound_identities, f"bound input file changed {phase}")
    observed_snapshots = tuple(
        _snapshot_tree(snapshot.root) for snapshot in preflight.tree_snapshots
    )
    _require(
        observed_snapshots == preflight.tree_snapshots,
        f"bound input tree universe changed {phase}",
    )
    return {
        "phase": phase,
        "bound_files_exact": True,
        "tree_universes_exact": True,
        "trees": [snapshot.public() for snapshot in observed_snapshots],
    }


def _capture_flat_output_universe(
    root: Path,
    expected_names: Sequence[str],
    *,
    forbidden_physical_aliases: Sequence[Path],
) -> tuple[dict[str, CapturedFile], dict[str, object]]:
    root = _lexical_absolute(root)
    _assert_plain_components(root)
    facts = os.lstat(root)
    _require(stat.S_ISDIR(facts.st_mode), f"output universe root is not a directory: {root}")
    canonical_root = root.resolve(strict=True)
    expected = sorted(expected_names)
    _require(
        len(expected) == len(set(expected))
        and all(Path(name).name == name and name not in {"", ".", ".."}
                for name in expected),
        "expected output universe contains an unsafe/non-unique flat name",
    )
    entries = sorted(canonical_root.iterdir(), key=lambda item: item.name)
    _require(
        [entry.name for entry in entries] == expected,
        "complete output artifact name universe drift",
    )
    captures: dict[str, CapturedFile] = {}
    for name in expected:
        lexical_path = canonical_root / name
        captured = _capture_file(lexical_path)
        _require(
            captured.path.parent == canonical_root and captured.path.name == name,
            f"final artifact escaped the flat output root: {name}",
        )
        captures[name] = captured
    ordered_names = sorted(captures)
    for index, name in enumerate(ordered_names):
        captured = captures[name]
        _require(
            captured.nlink == 1,
            f"final output artifact has a non-exclusive hardlink count: {name}",
        )
        for forbidden in forbidden_physical_aliases:
            try:
                aliases_input = os.path.samefile(captured.path, forbidden)
            except OSError as exc:
                raise DirectRawWitnessError(
                    f"cannot prove output/input physical distinctness: {name} vs {forbidden}"
                ) from exc
            _require(
                not aliases_input,
                f"final output artifact physically aliases a bound input: {name}",
            )
        for other_name in ordered_names[index + 1:]:
            try:
                aliases_output = os.path.samefile(
                    captured.path, captures[other_name].path,
                )
            except OSError as exc:
                raise DirectRawWitnessError(
                    f"cannot prove pairwise output distinctness: {name} vs {other_name}"
                ) from exc
            _require(
                not aliases_output,
                f"final output artifacts physically alias: {name} vs {other_name}",
            )
    members = [
        {
            "relative_path": name,
            "bytes": captures[name].bytes,
            "sha256": captures[name].sha256,
        }
        for name in expected
    ]
    return captures, {
        "schema": "phase8-complete-flat-output-manifest/v1",
        "scope": "every ordinary file present before detached terminal completion",
        "files": len(members),
        "bytes": sum(item["bytes"] for item in members),
        "canonical_members_sha256": hashlib.sha256(
            witness._canonical_bytes(members)
        ).hexdigest(),
        "members": members,
    }


def _declared_identity(
    identity: Mapping[str, object],
    *,
    root: Path,
    label: str,
) -> tuple[str, int, str]:
    raw_path = identity.get("path")
    _require(isinstance(raw_path, str), f"{label} declared path absent")
    path = Path(raw_path)
    _require(path.is_absolute(), f"{label} declared path is not absolute")
    _assert_plain_components(path)
    canonical = path.resolve(strict=True)
    _require(canonical.parent == root, f"{label} declared path escaped output root")
    size = identity.get("bytes")
    digest = identity.get("sha256")
    _require(type(size) is int and isinstance(digest, str), f"{label} identity shape drift")
    return canonical.name, size, digest


def _validate_complete_output_identities(
    captures: Mapping[str, CapturedFile],
    expected: Mapping[str, tuple[int, str]],
) -> None:
    _require(set(captures) == set(expected), "complete output identity key universe drift")
    for name, captured in captures.items():
        _require(
            (captured.bytes, captured.sha256) == expected[name],
            f"complete output artifact identity drift: {name}",
        )


TERMINAL_PRECONDITION_KEYS = {
    "post_result_input_trees_exact",
    "complete_output_artifact_universe_exact",
    "complete_output_identities_exact",
    "complete_output_members_physically_disjoint",
    "terminal_residue_exact",
    "output_root_still_disjoint",
}


def _require_terminal_preconditions(preconditions: Mapping[str, object]) -> None:
    checked = _require_keys(
        preconditions, TERMINAL_PRECONDITION_KEYS,
        "detached terminal completion preconditions",
    )
    _require(
        all(checked[key] is True for key in TERMINAL_PRECONDITION_KEYS),
        "detached terminal completion precondition failed",
    )


def _valid_sha256(value: object) -> bool:
    return (
        isinstance(value, str)
        and len(value) == 64
        and all(character in "0123456789abcdef" for character in value)
    )


def _validate_audit_code_contract(value: object) -> Mapping[str, object]:
    manifest = _require_keys(value, AUDIT_CODE_MANIFEST_KEYS, "completion audit-code manifest")
    _require(
        manifest.get("schema") == "phase8-direct-raw-audit-code/v1",
        "completion audit-code schema drift",
    )
    members = manifest.get("members")
    _require(isinstance(members, list), "completion audit-code members absent")
    normalized: list[Mapping[str, object]] = []
    observed_roles: set[str] = set()
    for index, raw_member in enumerate(members):
        member = _require_keys(
            raw_member, AUDIT_CODE_MEMBER_KEYS,
            f"completion audit-code member {index}",
        )
        role = member.get("role")
        _require(
            isinstance(role, str)
            and role in AUDIT_CODE_LOGICAL_PATHS
            and role not in observed_roles,
            f"completion audit-code role drift at member {index}",
        )
        _require(
            member.get("logical_path") == AUDIT_CODE_LOGICAL_PATHS[role]
            and type(member.get("bytes")) is int
            and int(member["bytes"]) >= 0
            and _valid_sha256(member.get("sha256")),
            f"completion audit-code identity drift: {role}",
        )
        observed_roles.add(role)
        normalized.append(member)
    _require(
        observed_roles == set(AUDIT_CODE_LOGICAL_PATHS)
        and type(manifest.get("roles")) is int
        and manifest.get("roles") == len(AUDIT_CODE_LOGICAL_PATHS)
        and members == sorted(members, key=lambda item: str(item["role"])),
        "completion audit-code role universe/order drift",
    )
    _require(
        manifest.get("canonical_members_sha256")
        == hashlib.sha256(witness._canonical_bytes(normalized)).hexdigest(),
        "completion audit-code aggregate digest drift",
    )
    return manifest


def _validate_complete_manifest_contract(value: object) -> tuple[Mapping[str, object], list[str]]:
    manifest = _require_keys(
        value, COMPLETE_OUTPUT_MANIFEST_KEYS,
        "detached completion complete-output manifest",
    )
    _require(
        manifest.get("schema") == "phase8-complete-flat-output-manifest/v1"
        and manifest.get("scope")
        == "every ordinary file present before detached terminal completion",
        "complete-output manifest schema/scope drift",
    )
    raw_members = manifest.get("members")
    _require(isinstance(raw_members, list), "complete-output members absent")
    members: list[Mapping[str, object]] = []
    names: list[str] = []
    for index, raw_member in enumerate(raw_members):
        member = _require_keys(
            raw_member, COMPLETE_OUTPUT_MEMBER_KEYS,
            f"complete-output member {index}",
        )
        name = member.get("relative_path")
        _require(
            isinstance(name, str)
            and Path(name).name == name
            and name not in {"", ".", ".."}
            and type(member.get("bytes")) is int
            and int(member["bytes"]) >= 0
            and _valid_sha256(member.get("sha256")),
            f"complete-output member identity drift: {index}",
        )
        names.append(name)
        members.append(member)
    _require(
        names == sorted(set(names)),
        "complete-output member name universe/order drift",
    )
    _require(
        type(manifest.get("files")) is int
        and type(manifest.get("bytes")) is int
        and int(manifest["files"]) >= 0
        and int(manifest["bytes"]) >= 0
        and manifest.get("files") == len(members)
        and manifest.get("bytes") == sum(int(item["bytes"]) for item in members)
        and manifest.get("canonical_members_sha256")
        == hashlib.sha256(witness._canonical_bytes(members)).hexdigest(),
        "complete-output manifest aggregate drift",
    )
    return manifest, names


def _validate_preterminal_product_claims(
    *,
    root: Path,
    preterminal: Mapping[str, object],
    manifest: Mapping[str, object],
    observed_chunks: Sequence[str],
    terminal_residue: Mapping[str, object],
) -> None:
    member_by_name = {
        str(item["relative_path"]): item for item in manifest["members"]
    }

    def bind_identity(value: object, expected_name: str, label: str) -> Mapping[str, object]:
        identity = _require_keys(value, IDENTITY_KEYS, label)
        name, size, digest = _declared_identity(identity, root=root, label=label)
        _require(name == expected_name, f"{label} filename drift")
        member = member_by_name.get(name)
        _require(
            isinstance(member, Mapping)
            and member.get("bytes") == size
            and member.get("sha256") == digest,
            f"{label} disagrees with complete-output manifest",
        )
        return identity

    result = _require_keys(
        preterminal.get("result"), PRODUCT_RESULT_KEYS,
        "preterminal product result summary",
    )
    _require(
        result.get("status") == "ok"
        and result.get("completion") == "complete"
        and result.get("verdict") in {"match", "diff"}
        and type(result.get("skipped_inputs")) is int
        and result.get("skipped_inputs") == 0
        and type(result.get("failed_inputs")) is int
        and result.get("failed_inputs") == 0
        and result.get("failures") == []
        and result.get("pairing_quality") == "exact"
        and isinstance(result.get("summary_lines"), list)
        and isinstance(result.get("warnings"), list)
        and _valid_sha256(result.get("pairing_trace_sha256"))
        and _valid_sha256(result.get("comparison_outcome_sha256")),
        "preterminal product success/pairing claims drift",
    )
    counts = _require_keys(
        result.get("counts"), PRODUCT_COUNT_KEYS,
        "preterminal product count summary",
    )
    _require(counts.get("known") is True, "preterminal product counts are not known")
    for key in PRODUCT_COUNT_KEYS - {"known", "per_field_counts"}:
        _require(
            type(counts.get(key)) is int and int(counts[key]) >= 0,
            f"preterminal product count drift: {key}",
        )
    _require(
        isinstance(counts.get("per_field_counts"), Mapping),
        "preterminal per-field counts absent",
    )
    _require(
        type(result.get("pairing_trace_count")) is int
        and type(result.get("duplicate_group_count")) is int
        and result.get("pairing_trace_count") == result.get("duplicate_group_count")
        and int(result["pairing_trace_count"]) >= 0,
        "preterminal pairing trace census drift",
    )

    outputs = _require_keys(preterminal.get("outputs"), {"formulas", "values"}, "workbooks")
    sidecars = _require_keys(
        preterminal.get("outcome_sidecars"), {"formulas", "values"},
        "outcome sidecars",
    )
    output_names = {
        "formulas": "comparison.xlsx",
        "values": "comparison (values).xlsx",
    }
    sidecar_names = {
        flavor: f"{name}.outcome.json" for flavor, name in output_names.items()
    }
    for flavor in ("formulas", "values"):
        bind_identity(outputs[flavor], output_names[flavor], f"{flavor} workbook")
        bind_identity(sidecars[flavor], sidecar_names[flavor], f"{flavor} outcome sidecar")

    generation = _require_keys(
        result.get("artifact_generation"), ARTIFACT_GENERATION_KEYS,
        "product artifact generation",
    )
    _require(
        generation.get("completion") == "complete"
        and generation.get("publication_state") == "committed"
        and generation.get("requested_mode") == "both",
        "product artifact-generation state drift",
    )
    generation_members = generation.get("members")
    _require(
        isinstance(generation_members, list) and len(generation_members) == 2,
        "product generation member census drift",
    )
    generation_by_flavor: dict[str, Mapping[str, object]] = {}
    for index, raw_member in enumerate(generation_members):
        member = _require_keys(
            raw_member, GENERATION_MEMBER_KEYS, f"generation member {index}",
        )
        flavor = member.get("flavor")
        _require(
            flavor in {"formulas", "values"} and flavor not in generation_by_flavor,
            f"generation member flavor drift: {flavor!r}",
        )
        generation_by_flavor[str(flavor)] = member
    for flavor, expected_role in {"formulas": "best_effort", "values": "canonical"}.items():
        identity = outputs[flavor]
        member = generation_by_flavor[flavor]
        _require(
            member.get("commit_role") == expected_role
            and member.get("path") == identity["path"]
            and member.get("bytes") == identity["bytes"]
            and member.get("sha256") == identity["sha256"],
            f"generation/{flavor} workbook identity drift",
        )
    persisted = _require_keys(
        result.get("persisted_members"), {"formulas", "values"},
        "persisted comparison members",
    )
    for flavor in ("formulas", "values"):
        item = _require_keys(
            persisted[flavor], PERSISTED_MEMBER_KEYS,
            f"persisted {flavor} comparison member",
        )
        _require(
            item.get("trusted") is True
            and item.get("current") is True
            and item.get("completion") == "complete"
            and isinstance(item.get("source"), str),
            f"persisted {flavor} comparison member truth drift",
        )

    decoded = _require_keys(
        preterminal.get("decoded_comparison_payload"), DECODED_PAYLOAD_KEYS,
        "decoded comparison payload",
    )
    _require(
        decoded.get("product_reader_decoded_both_peers") is True
        and decoded.get("independent_decode_exact") is True
        and decoded.get("comparison_schema_version") == 3
        and decoded.get("payload_schema_version") == 1
        and decoded.get("encoding") == "canonical-json-zlib-chunks-v1"
        and type(decoded.get("decoded_bytes")) is int
        and int(decoded["decoded_bytes"]) > 0
        and _valid_sha256(decoded.get("decoded_sha256")),
        "decoded comparison payload aggregate drift",
    )
    raw_chunks = decoded.get("chunks")
    _require(isinstance(raw_chunks, list) and raw_chunks, "decoded payload chunks absent")
    chunks: list[Mapping[str, object]] = []
    for index, raw_chunk in enumerate(raw_chunks):
        chunk = _require_keys(raw_chunk, DECODED_CHUNK_KEYS, f"decoded chunk {index}")
        name = chunk.get("relative_path")
        _require(
            chunk.get("index") == index
            and isinstance(name, str)
            and witness.PAYLOAD_BASENAME_RE.fullmatch(name) is not None
            and type(chunk.get("compressed_bytes")) is int
            and int(chunk["compressed_bytes"]) > 0
            and _valid_sha256(chunk.get("compressed_sha256"))
            and type(chunk.get("decoded_bytes")) is int
            and int(chunk["decoded_bytes"]) > 0
            and _valid_sha256(chunk.get("decoded_sha256")),
            f"decoded payload chunk claim drift: {index}",
        )
        member = member_by_name.get(str(name))
        _require(
            isinstance(member, Mapping)
            and member.get("bytes") == chunk.get("compressed_bytes")
            and member.get("sha256") == chunk.get("compressed_sha256"),
            f"decoded payload chunk/manifest identity drift: {name}",
        )
        chunks.append(chunk)
    _require(
        [str(item["relative_path"]) for item in chunks] == list(observed_chunks)
        and sum(int(item["decoded_bytes"]) for item in chunks)
        == decoded.get("decoded_bytes"),
        "decoded payload chunk universe/byte total drift",
    )

    publication_artifacts = _require_keys(
        preterminal.get("publication_artifacts"), PUBLICATION_ARTIFACT_KEYS,
        "preterminal publication artifacts",
    )
    _require(
        publication_artifacts.get("payload_chunks") == raw_chunks
        and publication_artifacts.get("outcome_sidecars") == sidecars
        and publication_artifacts.get("permanent_lease")
        == terminal_residue.get("permanent_lease"),
        "preterminal publication artifact claims disagree",
    )
    bind_identity(
        preterminal.get("product_code_manifest"), PRODUCT_CODE_MANIFEST_NAME,
        "product-code manifest",
    )
    bind_identity(
        preterminal.get("artifact_manifest"), ARTIFACT_MANIFEST_NAME,
        "artifact manifest",
    )


def _validate_terminal_completion_contract(
    path: Path,
    payload: Mapping[str, object],
    preconditions: Mapping[str, object],
    *,
    expected_leg: str,
    expected_audit_code_manifest: Mapping[str, object],
) -> None:
    _require(expected_leg in LEG_CHOICES, "expected completion leg is invalid")
    expected_audit_code = _validate_audit_code_contract(
        expected_audit_code_manifest
    )
    completion = _require_keys(payload, COMPLETION_KEYS, "detached terminal completion")
    _require(
        completion.get("schema_version") == 1
        and completion.get("audit") == COMPLETION_AUDIT
        and completion.get("status") == COMPLETION_STATUS
        and completion.get("terminal") is True
        and completion.get("artifact_status") == ARTIFACT_STATUS
        and completion.get("acceptance_eligible") is False
        and completion.get("stage8_family_accepted") is False,
        "detached terminal completion status/nonacceptance contract drift",
    )
    leg = completion.get("leg")
    _require(
        leg == expected_leg,
        "detached completion leg is not the externally requested exact leg",
    )
    _require(
        completion.get("output_root") == str(path.parent),
        "detached completion output-root binding drift",
    )
    checked_preconditions = _require_keys(
        completion.get("terminal_preconditions"), TERMINAL_PRECONDITION_KEYS,
        "completion-embedded terminal preconditions",
    )
    _require(
        dict(checked_preconditions) == dict(preconditions),
        "completion/argument terminal preconditions disagree",
    )
    _require_terminal_preconditions(checked_preconditions)
    invariants = _require_keys(
        completion.get("invariants"), COMPLETION_INVARIANT_KEYS,
        "detached completion invariants",
    )
    _require(all(invariants[key] is True for key in invariants), "completion invariant false")
    audit_code = _validate_audit_code_contract(completion.get("audit_code"))
    _require(
        audit_code == expected_audit_code,
        "completion audit-code ledger is not the freshly captured physical ledger",
    )
    manifest, member_names = _validate_complete_manifest_contract(
        completion.get("complete_output_artifact_manifest")
    )
    expected_final_names = completion.get("expected_final_artifact_names")
    _require(
        isinstance(expected_final_names, list)
        and expected_final_names == sorted(set(expected_final_names))
        and expected_final_names == sorted([*member_names, COMPLETION_NAME])
        and COMPLETION_NAME not in member_names,
        "detached completion final/self-excluded member universe drift",
    )
    required_fixed_names = {
        "comparison.xlsx", "comparison (values).xlsx",
        "comparison.xlsx.outcome.json", "comparison (values).xlsx.outcome.json",
        witness.PUBLICATION_LEASE_NAME, PRODUCT_CODE_MANIFEST_NAME,
        ARTIFACT_MANIFEST_NAME, RESULT_NAME,
    }
    _require(
        required_fixed_names <= set(member_names),
        "detached completion fixed production artifact universe drift",
    )
    observed_chunks = sorted(
        name for name in member_names if witness.PAYLOAD_BASENAME_RE.fullmatch(name)
    )
    _require(
        completion.get("payload_chunk_names") == observed_chunks
        and bool(observed_chunks)
        and set(member_names) == required_fixed_names | set(observed_chunks),
        "detached completion payload-chunk universe drift",
    )
    identity = _require_keys(
        completion.get("preterminal_result"), IDENTITY_KEYS,
        "detached completion preterminal-result identity",
    )
    result_name, result_bytes, result_sha = _declared_identity(
        identity, root=path.parent, label="detached completion preterminal result",
    )
    _require(result_name == RESULT_NAME, "completion preterminal result name drift")
    captured_result = _capture_file(path.parent / RESULT_NAME)
    _require(
        (captured_result.bytes, captured_result.sha256) == (result_bytes, result_sha),
        "completion preterminal-result byte identity drift",
    )
    preterminal = _require_keys(
        _strict_json_bytes(captured_result.payload, RESULT_NAME, canonical=True),
        PRETERMINAL_KEYS, "preterminal result",
    )
    _require(
        preterminal.get("schema_version") == 1
        and preterminal.get("audit") == PRETERMINAL_AUDIT
        and preterminal.get("status") == PRETERMINAL_STATUS
        and preterminal.get("terminal") is False
        and preterminal.get("artifact_status") == ARTIFACT_STATUS
        and preterminal.get("acceptance_eligible") is False
        and preterminal.get("stage8_family_accepted") is False
        and preterminal.get("leg") == leg
        and preterminal.get("output_root") == str(path.parent)
        and preterminal.get("required_detached_terminal_completion") == COMPLETION_NAME,
        "preterminal result status/nonacceptance/binding contract drift",
    )
    _require(
        isinstance(preterminal.get("reason"), str) and bool(preterminal["reason"]),
        "preterminal result reason absent",
    )
    preterminal_invariants = _require_keys(
        preterminal.get("invariants"), PRETERMINAL_INVARIANT_KEYS,
        "preterminal result invariants",
    )
    _require(
        all(preterminal_invariants[key] is True for key in preterminal_invariants),
        "preterminal result invariant false",
    )
    _require(
        preterminal.get("expected_precompletion_artifact_names") == member_names,
        "preterminal/complete-output artifact universes disagree",
    )
    _require(
        preterminal.get("audit_code") == audit_code
        and preterminal.get("audit_code_mutations")
        == completion.get("audit_code_mutations")
        and preterminal.get("output_containment_mutations")
        == completion.get("output_containment_mutations")
        and preterminal.get("publication_lifecycle_mutations")
        == completion.get("publication_lifecycle_mutations"),
        "completion/preterminal audit-control bindings disagree",
    )
    post_result = _require_keys(
        completion.get("post_result_input_revalidation"), INPUT_REVALIDATION_KEYS,
        "completion post-result input revalidation",
    )
    _require(
        post_result.get("bound_files_exact") is True
        and post_result.get("tree_universes_exact") is True
        and post_result.get("phase") == "after nonterminal result publication"
        and isinstance(post_result.get("trees"), list),
        "completion post-result input revalidation truth drift",
    )
    direct_twin_summary = preterminal.get("direct_twin_preimport_validation")
    _require(
        isinstance(direct_twin_summary, Mapping)
        and isinstance(direct_twin_summary.get("frozen_input_trees"), list)
        and bool(direct_twin_summary["frozen_input_trees"])
        and post_result.get("trees") == direct_twin_summary["frozen_input_trees"],
        "completion post-result tree ledger is not the frozen preimport baseline",
    )
    audit_mutations = _require_keys(
        completion.get("audit_code_mutations"), AUDIT_CODE_MUTATION_KEYS,
        "completion audit-code mutations",
    )
    expected_audit_roles = sorted(AUDIT_CODE_LOGICAL_PATHS)
    _require(
        audit_mutations.get("roles_copied") == expected_audit_roles
        and audit_mutations.get("mutated_roles") == expected_audit_roles
        and audit_mutations.get("unchanged_copied_baseline_passed") is True
        and audit_mutations.get("every_copied_role_byte_mutation_rejected") is True
        and audit_mutations.get("reported_role_reparse_redirections_rejected")
        == expected_audit_roles
        and audit_mutations.get(
            "every_reported_role_reparse_redirection_rejected"
        ) is True
        and audit_mutations.get("physical_logical_paths_exact") is True
        and audit_mutations.get("real_audit_code_unchanged") is True,
        "completion audit-code mutation truth drift",
    )
    containment = _require_keys(
        completion.get("output_containment_mutations"), CONTAINMENT_MUTATION_KEYS,
        "completion containment mutations",
    )
    _require(
        all(
            containment.get(key) is True
            for key in {
                "child_of_input_rejected", "hardlink_alias_rejected",
                "lexical_alias_rejected", "parent_of_input_rejected",
                "same_root_rejected", "valid_sibling_accepted",
            }
        )
        and all(
            containment.get(key) == "executed_and_rejected"
            for key in {
                "broken_link_occupant_mutation", "directory_link_mutation",
                "file_link_mutation",
            }
        ),
        "completion containment mutation truth drift",
    )
    publication = _require_keys(
        completion.get("publication_lifecycle_mutations"), PUBLICATION_MUTATION_KEYS,
        "completion publication controls",
    )
    fixture_only = publication.get("disposable_control_only")
    _require(type(fixture_only) is bool, "publication control fixture marker drift")
    if fixture_only:
        _require(
            path.parent.parent.name.startswith("phase8-direct-publication-controls-")
            and str(preterminal.get("reason", "")).startswith("DISPOSABLE CONTROL ONLY"),
            "nonauthoritative publication fixture escaped its disposable temp scope",
        )
    required_publication_truths = {
        "all_publication_controls_passed",
        "all_failed_preconditions_left_no_terminal_file",
        "bound_input_hardlink_rejected",
        "completion_contract_baseline_accepted",
        "completion_semantic_mutations_left_no_terminal_file",
        "destination_collision_rejected_without_overwrite",
        "extra_name_universe_rejected", "missing_name_universe_rejected",
        "pairwise_output_hardlink_rejected",
        "second_callback_artifact_mutation_rejected",
        "staging_payload_mutation_rejected", "staging_residue_rejected",
        "unchanged_baseline_accepted",
    }
    publication_roles = publication.get("post_manifest_artifact_roles_rejected")
    publication_role_set = set(publication_roles) if isinstance(publication_roles, list) else set()
    publication_role_chunks = {
        name for name in publication_role_set
        if isinstance(name, str) and witness.PAYLOAD_BASENAME_RE.fullmatch(name)
    }
    _require(
        all(publication.get(key) is True for key in required_publication_truths)
        and publication.get("final_artifact_symlink_mutation")
        == "executed_and_rejected"
        and set(publication.get("failed_terminal_preconditions", []))
        == TERMINAL_PRECONDITION_KEYS
        and set(publication.get("completion_semantic_mutations_rejected", []))
        == COMPLETION_SEMANTIC_MUTATION_LABELS
        and publication_role_set == required_fixed_names | publication_role_chunks
        and len(publication_role_chunks) == 1,
        "completion publication mutation truth drift",
    )
    publication_success = _require_keys(
        publication.get("successful_terminal_control"),
        SUCCESSFUL_TERMINAL_CONTROL_KEYS,
        "completion successful terminal control",
    )
    _require(
        publication_success.get("exit_zero") is True
        and publication_success.get("precommit_stdout_nonterminal") is True
        and publication_success.get("exact_final_universe") is True
        and publication_success.get("no_pending_residue") is True
        and publication_success.get("completion_canonical_and_exact") is True
        and publication_success.get("validation_phases")
        == ["after_completion_staging", "immediately_before_terminal_commit"],
        "completion positive publication control truth drift",
    )
    residue = _require_keys(
        completion.get("terminal_residue_gate"), RESIDUE_KEYS,
        "completion terminal residue gate",
    )
    residue_universe = _require_keys(
        residue.get("exact_artifact_universe"), RESIDUE_UNIVERSE_KEYS,
        "completion terminal residue universe",
    )
    core_names = {
        "comparison.xlsx", "comparison (values).xlsx",
        "comparison.xlsx.outcome.json", "comparison (values).xlsx.outcome.json",
        witness.PUBLICATION_LEASE_NAME,
    }
    audit_names = {PRODUCT_CODE_MANIFEST_NAME, ARTIFACT_MANIFEST_NAME, RESULT_NAME}
    _require(
        residue.get("transient_residue") == []
        and residue_universe.get("core_names") == sorted(core_names)
        and residue_universe.get("audit_names") == sorted(audit_names)
        and residue_universe.get("payload_chunks") == observed_chunks
        and isinstance(residue.get("rejected_name_classes"), list)
        and bool(residue.get("rejected_name_classes")),
        "completion terminal residue truth drift",
    )
    lease = _require_keys(
        residue.get("permanent_lease"), IDENTITY_KEYS,
        "completion permanent lease identity",
    )
    lease_name, lease_bytes, lease_sha = _declared_identity(
        lease, root=path.parent, label="completion permanent lease",
    )
    member_by_name = {
        str(item["relative_path"]): item for item in manifest["members"]
    }
    _require(
        lease_name == witness.PUBLICATION_LEASE_NAME
        and lease_bytes == 0
        and lease_sha == hashlib.sha256(b"").hexdigest()
        and member_by_name[lease_name]["bytes"] == lease_bytes
        and member_by_name[lease_name]["sha256"] == lease_sha,
        "completion permanent lease truth/manifest binding drift",
    )
    lease_exception = residue.get("permanent_lease_exception")
    _require(
        isinstance(lease_exception, Mapping)
        and lease_exception.get("relative_path") == witness.PUBLICATION_LEASE_NAME
        and type(lease_exception.get("required_bytes")) is int
        and lease_exception.get("required_bytes") == 0,
        "completion permanent lease exception drift",
    )
    _validate_preterminal_product_claims(
        root=path.parent,
        preterminal=preterminal,
        manifest=manifest,
        observed_chunks=observed_chunks,
        terminal_residue=residue,
    )


def _publish_terminal_completion(
    path: Path,
    payload: Mapping[str, object],
    *,
    preconditions: Mapping[str, object],
    expected_leg: str,
    expected_audit_code_manifest: Mapping[str, object],
    forbidden_physical_aliases: Sequence[Path],
    final_validation: Callable[[str], None] | None = None,
) -> int:
    """Commit the detached PASS as the last fallible action on Windows.

    The candidate is completely serialized, fsynced, recaptured, and reported
    before the final same-directory rename.  Nothing performs filesystem or
    stdout work after the rename.
    """
    _require_terminal_preconditions(preconditions)
    _validate_terminal_completion_contract(
        path,
        payload,
        preconditions,
        expected_leg=expected_leg,
        expected_audit_code_manifest=expected_audit_code_manifest,
    )
    _require(final_validation is not None, "terminal final-validation callback absent")
    _require(os.name == "nt", "terminal exclusive-rename contract is Windows-only")
    _require(path.is_absolute(), "terminal completion path must be absolute")
    _assert_plain_components(path.parent)
    _require(not os.path.lexists(path), f"terminal completion already exists: {path}")
    raw = witness._canonical_bytes(payload, newline=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".pending", dir=path.parent,
    )
    temporary = Path(temporary_name)
    committed = False
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(raw)
            handle.flush()
            os.fsync(handle.fileno())
        captured = _capture_file(temporary)
        _require(
            captured.bytes == len(raw)
            and captured.sha256 == hashlib.sha256(raw).hexdigest()
            and captured.payload == raw,
            "detached terminal completion staging identity drift",
        )
        expected_final_names = payload.get("expected_final_artifact_names")
        _require(
            isinstance(expected_final_names, list)
            and len(expected_final_names) == len(set(expected_final_names))
            and path.name in expected_final_names,
            "detached completion final artifact declaration drift",
        )
        expected_staging_names = sorted(
            (set(expected_final_names) - {path.name}) | {temporary.name}
        )
        final_validation("after_completion_staging")
        staged_captures, _staged_manifest = _capture_flat_output_universe(
            path.parent, expected_staging_names,
            forbidden_physical_aliases=forbidden_physical_aliases,
        )
        declared_manifest = _require_keys(
            payload.get("complete_output_artifact_manifest"),
            {
                "schema", "scope", "files", "bytes",
                "canonical_members_sha256", "members",
            },
            "detached completion complete-output manifest",
        )
        declared_members = declared_manifest.get("members")
        _require(isinstance(declared_members, list), "complete-output members absent")
        declared_identities: dict[str, tuple[int, str]] = {}
        for index, member in enumerate(declared_members):
            member = _require_keys(
                member, {"relative_path", "bytes", "sha256"},
                f"complete-output member {index}",
            )
            name = member["relative_path"]
            _require(
                isinstance(name, str) and Path(name).name == name,
                f"unsafe complete-output member name: {name!r}",
            )
            _require(name not in declared_identities, f"duplicate complete-output member: {name}")
            declared_identities[name] = (int(member["bytes"]), str(member["sha256"]))
        _require(
            set(declared_identities) == set(expected_final_names) - {path.name},
            "detached completion member/final-name universes disagree",
        )
        _validate_complete_output_identities(
            {
                name: item for name, item in staged_captures.items()
                if name != temporary.name
            },
            declared_identities,
        )
        staged_again = staged_captures[temporary.name]
        _require(
            (staged_again.bytes, staged_again.sha256)
            == (captured.bytes, captured.sha256),
            "terminal completion staging file changed during final universe recapture",
        )
        final_validation("immediately_before_terminal_commit")
        final_captures, _final_manifest = _capture_flat_output_universe(
            path.parent, expected_staging_names,
            forbidden_physical_aliases=forbidden_physical_aliases,
        )
        _validate_complete_output_identities(
            {
                name: item for name, item in final_captures.items()
                if name != temporary.name
            },
            declared_identities,
        )
        staged_final = final_captures[temporary.name]
        _require(
            (staged_final.bytes, staged_final.sha256, staged_final.token)
            == (captured.bytes, captured.sha256, captured.token),
            "terminal completion staging identity changed before commit",
        )
        _assert_plain_components(path.parent)
        _require(
            not os.path.lexists(path),
            "terminal completion destination appeared before commit",
        )
        print(json.dumps({
            "status": "TERMINAL_COMPLETION_PREPARED_NOT_COMMITTED",
            "terminal": False,
            "terminal_completion_ready": {
                "path": str(path),
                "bytes": captured.bytes,
                "sha256": captured.sha256,
            },
        }, ensure_ascii=False, sort_keys=True), flush=True)
        # On the required Windows runtime os.rename is same-volume atomic and
        # refuses to replace an existing destination.  This is deliberately
        # the final fallible operation in the successful process.
        os.rename(temporary, path)
        committed = True
        return 0
    finally:
        if not committed and os.path.lexists(temporary):
            try:
                temporary.unlink()
            except OSError:
                pass


def _run_publication_mutation_probes(
    *,
    audit_code_manifest: Mapping[str, object],
    audit_code_mutations: Mapping[str, object],
    containment_mutations: Mapping[str, object],
) -> dict[str, object]:
    with tempfile.TemporaryDirectory(
        prefix="phase8-direct-publication-controls-", dir=VISUAL_ROOT,
    ) as temporary:
        base = Path(temporary).resolve(strict=True)
        all_true = {key: True for key in TERMINAL_PRECONDITION_KEYS}
        semantic_mutation_labels = sorted(COMPLETION_SEMANTIC_MUTATION_LABELS)
        _validate_audit_code_contract(audit_code_manifest)
        publication_control_fixture = {
            "all_publication_controls_passed": True,
            "disposable_control_only": True,
            "unchanged_baseline_accepted": True,
            "bound_input_hardlink_rejected": True,
            "pairwise_output_hardlink_rejected": True,
            "missing_name_universe_rejected": True,
            "extra_name_universe_rejected": True,
            "second_callback_artifact_mutation_rejected": True,
            "staging_payload_mutation_rejected": True,
            "staging_residue_rejected": True,
            "destination_collision_rejected_without_overwrite": True,
            "final_artifact_symlink_mutation": "executed_and_rejected",
            "all_failed_preconditions_left_no_terminal_file": True,
            "completion_semantic_mutations_left_no_terminal_file": True,
            "completion_semantic_mutations_rejected": semantic_mutation_labels,
            "completion_contract_baseline_accepted": True,
            "failed_terminal_preconditions": sorted(TERMINAL_PRECONDITION_KEYS),
            "successful_terminal_control": {
                "completion_canonical_and_exact": True,
                "exact_final_universe": True,
                "exit_zero": True,
                "no_pending_residue": True,
                "precommit_stdout_nonterminal": True,
                "validation_phases": [
                    "after_completion_staging",
                    "immediately_before_terminal_commit",
                ],
            },
        }
        role_payloads = {
            "comparison.xlsx": b"formulas-workbook-baseline",
            "comparison (values).xlsx": b"values-workbook-baseline",
            "comparison.xlsx.outcome.json": b"formulas-outcome-sidecar-baseline",
            "comparison (values).xlsx.outcome.json": b"values-outcome-sidecar-baseline",
            (
                ".cmpv3-" + "0" * 64 + "-000001-" + "1" * 64
                + ".comparison-payload.zlib"
            ): b"payload-chunk-baseline",
            witness.PUBLICATION_LEASE_NAME: b"",
            PRODUCT_CODE_MANIFEST_NAME: b"product-code-manifest-baseline",
            ARTIFACT_MANIFEST_NAME: b"artifact-manifest-baseline",
            RESULT_NAME: b"nonterminal-result-baseline",
        }
        publication_control_fixture["post_manifest_artifact_roles_rejected"] = sorted(
            role_payloads
        )

        def prepare(
            label: str,
        ) -> tuple[
            Path, dict[str, Path], dict[str, CapturedFile],
            dict[str, object], dict[str, tuple[int, str]], dict[str, object],
        ]:
            output = base / label
            output.mkdir()
            files: dict[str, Path] = {}
            for name, raw in role_payloads.items():
                path = output / name
                path.write_bytes(raw)
                files[name] = path
            names = sorted(files)
            initial = {
                name: _capture_file(path)
                for name, path in files.items() if name != RESULT_NAME
            }
            chunk_names = sorted(
                name for name in names if witness.PAYLOAD_BASENAME_RE.fullmatch(name)
            )
            _require(len(chunk_names) == 1, "disposable control chunk census drift")
            chunk_name = chunk_names[0]
            decoded_chunks = [{
                "index": 0,
                "relative_path": chunk_name,
                "compressed_bytes": initial[chunk_name].bytes,
                "compressed_sha256": initial[chunk_name].sha256,
                "decoded_bytes": 1,
                "decoded_sha256": "2" * 64,
            }]
            workbooks = {
                "formulas": initial["comparison.xlsx"].identity(),
                "values": initial["comparison (values).xlsx"].identity(),
            }
            sidecars = {
                "formulas": initial["comparison.xlsx.outcome.json"].identity(),
                "values": initial["comparison (values).xlsx.outcome.json"].identity(),
            }
            product_result = {
                "status": "ok",
                "completion": "complete",
                "verdict": "match",
                "skipped_inputs": 0,
                "failed_inputs": 0,
                "summary_lines": ["DISPOSABLE CONTROL ONLY"],
                "counts": {
                    "known": True,
                    "paired_rows": 1,
                    "side_a_only_rows": 0,
                    "side_b_only_rows": 0,
                    "differing_rows": 0,
                    "differing_cells": 0,
                    "asserted_cells": 1,
                    "context_cells": 0,
                    "per_field_counts": {},
                },
                "warnings": [],
                "failures": [],
                "pairing_trace_count": 0,
                "duplicate_group_count": 0,
                "pairing_quality": "exact",
                "pairing_trace_sha256": "3" * 64,
                "comparison_outcome_sha256": "4" * 64,
                "artifact_generation": {
                    "completion": "complete",
                    "publication_state": "committed",
                    "requested_mode": "both",
                    "members": [
                        {
                            "flavor": flavor,
                            "commit_role": (
                                "best_effort" if flavor == "formulas" else "canonical"
                            ),
                            "path": workbooks[flavor]["path"],
                            "bytes": workbooks[flavor]["bytes"],
                            "sha256": workbooks[flavor]["sha256"],
                        }
                        for flavor in ("formulas", "values")
                    ],
                },
                "persisted_members": {
                    flavor: {
                        "trusted": True,
                        "current": True,
                        "completion": "complete",
                        "source": "disposable-control-only",
                    }
                    for flavor in ("formulas", "values")
                },
            }
            decoded_payload = {
                "product_reader_decoded_both_peers": True,
                "independent_decode_exact": True,
                "comparison_schema_version": 3,
                "payload_schema_version": 1,
                "encoding": "canonical-json-zlib-chunks-v1",
                "decoded_bytes": 1,
                "decoded_sha256": "5" * 64,
                "chunks": decoded_chunks,
            }
            lease_identity = initial[witness.PUBLICATION_LEASE_NAME].identity()
            preterminal = {
                "schema_version": 1,
                "audit": PRETERMINAL_AUDIT,
                "status": PRETERMINAL_STATUS,
                "terminal": False,
                "artifact_status": ARTIFACT_STATUS,
                "acceptance_eligible": False,
                "stage8_family_accepted": False,
                "reason": "DISPOSABLE CONTROL ONLY; never family acceptance.",
                "leg": "excel_vs_raw_tsn",
                "output_root": str(output),
                "direct_twin_preimport_validation": {
                    "disposable_control_only": True,
                    "frozen_input_trees": [{"disposable_control_only": True}],
                },
                "audit_code": audit_code_manifest,
                "audit_code_mutations": audit_code_mutations,
                "output_containment_mutations": containment_mutations,
                "publication_lifecycle_mutations": publication_control_fixture,
                "input_tree_revalidation": {"disposable_control_only": True},
                "result": product_result,
                "outputs": workbooks,
                "outcome_sidecars": sidecars,
                "decoded_comparison_payload": decoded_payload,
                "publication_artifacts": {
                    "payload_chunks": decoded_chunks,
                    "outcome_sidecars": sidecars,
                    "permanent_lease": lease_identity,
                },
                "residue_gate": {"disposable_control_only": True},
                "loaded_product_code": {"disposable_control_only": True},
                "product_code_manifest": initial[PRODUCT_CODE_MANIFEST_NAME].identity(),
                "artifact_manifest_before_result": {"disposable_control_only": True},
                "artifact_manifest": initial[ARTIFACT_MANIFEST_NAME].identity(),
                "deterministic_serialization": {"disposable_control_only": True},
                "expected_precompletion_artifact_names": names,
                "required_detached_terminal_completion": COMPLETION_NAME,
                "invariants": {
                    key: True for key in PRETERMINAL_INVARIANT_KEYS
                },
            }
            files[RESULT_NAME].write_bytes(
                witness._canonical_bytes(preterminal, newline=True)
            )
            captures, manifest = _capture_flat_output_universe(
                output, names, forbidden_physical_aliases=(),
            )
            expected = {
                name: (captures[name].bytes, captures[name].sha256)
                for name in names
            }
            _validate_complete_output_identities(captures, expected)
            payload = {
                "schema_version": 1,
                "audit": COMPLETION_AUDIT,
                "status": COMPLETION_STATUS,
                "terminal": True,
                "artifact_status": ARTIFACT_STATUS,
                "acceptance_eligible": False,
                "stage8_family_accepted": False,
                "leg": "excel_vs_raw_tsn",
                "output_root": str(output),
                "preterminal_result": captures[RESULT_NAME].identity(),
                "audit_code": audit_code_manifest,
                "audit_code_mutations": audit_code_mutations,
                "post_result_input_revalidation": {
                    "bound_files_exact": True,
                    "phase": "after nonterminal result publication",
                    "tree_universes_exact": True,
                    "trees": [{"disposable_control_only": True}],
                },
                "expected_final_artifact_names": sorted([*names, COMPLETION_NAME]),
                "complete_output_artifact_manifest": manifest,
                "terminal_residue_gate": {
                    "transient_residue": [],
                    "exact_artifact_universe": {
                        "core_names": sorted({
                            "comparison.xlsx", "comparison (values).xlsx",
                            "comparison.xlsx.outcome.json",
                            "comparison (values).xlsx.outcome.json",
                            witness.PUBLICATION_LEASE_NAME,
                        }),
                        "audit_names": sorted({
                            PRODUCT_CODE_MANIFEST_NAME,
                            ARTIFACT_MANIFEST_NAME,
                            RESULT_NAME,
                        }),
                        "payload_chunks": sorted(
                            name for name in names
                            if witness.PAYLOAD_BASENAME_RE.fullmatch(name)
                        ),
                    },
                    "permanent_lease":
                        captures[witness.PUBLICATION_LEASE_NAME].identity(),
                    "permanent_lease_exception": {
                        "relative_path": witness.PUBLICATION_LEASE_NAME,
                        "required_bytes": 0,
                    },
                    "rejected_name_classes": ["disposable control exact universe"],
                },
                "payload_chunk_names": sorted(
                    name for name in names
                    if witness.PAYLOAD_BASENAME_RE.fullmatch(name)
                ),
                "terminal_preconditions": all_true,
                "output_containment_mutations": containment_mutations,
                "publication_lifecycle_mutations": publication_control_fixture,
                "invariants": {
                    key: True for key in COMPLETION_INVARIANT_KEYS
                },
            }
            return output, files, captures, manifest, expected, payload

        def pending_names(output: Path) -> list[str]:
            return sorted(
                item.name for item in output.iterdir()
                if item.name.startswith(f".{COMPLETION_NAME}.")
                and item.name.endswith(".pending")
            )

        def require_no_terminal(output: Path, label: str) -> None:
            _require(
                not os.path.lexists(output / COMPLETION_NAME),
                f"{label} left an authoritative terminal completion",
            )
            _require(not pending_names(output), f"{label} left terminal staging residue")

        (
            role_output, role_files, role_captures, _role_manifest,
            role_expected, _role_completion,
        ) = prepare("artifact-role-identity-controls")
        _validate_complete_output_identities(role_captures, role_expected)
        unchanged_baseline_accepted = True
        rejected_artifact_roles: list[str] = []
        for name in sorted(role_files):
            original = role_captures[name].payload
            role_files[name].write_bytes(original + b"-same-name-mutation")
            mutated, _mutated_manifest = _capture_flat_output_universe(
                role_output, sorted(role_files), forbidden_physical_aliases=(),
            )
            rejected = False
            try:
                _validate_complete_output_identities(mutated, role_expected)
            except DirectRawWitnessError:
                rejected = True
            _require(rejected, f"post-manifest artifact-class mutation escaped: {name}")
            rejected_artifact_roles.append(name)
            role_files[name].write_bytes(original)
            restored, _restored_manifest = _capture_flat_output_universe(
                role_output, sorted(role_files), forbidden_physical_aliases=(),
            )
            _validate_complete_output_identities(restored, role_expected)

        missing_output, missing_files, *_missing = prepare("missing-name-control")
        missing_name = sorted(missing_files)[0]
        missing_files[missing_name].unlink()
        missing_name_rejected = False
        try:
            _capture_flat_output_universe(
                missing_output, sorted(missing_files), forbidden_physical_aliases=(),
            )
        except DirectRawWitnessError:
            missing_name_rejected = True
        _require(missing_name_rejected, "missing final artifact name escaped universe gate")

        extra_output, extra_files, *_extra = prepare("extra-name-control")
        (extra_output / "undeclared-extra.bin").write_bytes(b"extra")
        extra_name_rejected = False
        try:
            _capture_flat_output_universe(
                extra_output, sorted(extra_files), forbidden_physical_aliases=(),
            )
        except DirectRawWitnessError:
            extra_name_rejected = True
        _require(extra_name_rejected, "extra final artifact name escaped universe gate")

        (
            semantic_output, semantic_files, semantic_captures,
            _semantic_manifest, _semantic_expected, semantic_payload,
        ) = prepare("completion-semantic-mutation-controls")
        _validate_terminal_completion_contract(
            semantic_output / COMPLETION_NAME, semantic_payload, all_true,
            expected_leg="excel_vs_raw_tsn",
            expected_audit_code_manifest=audit_code_manifest,
        )
        completion_contract_baseline_accepted = True
        semantic_result_baseline = semantic_captures[RESULT_NAME].payload
        rejected_completion_semantics: list[str] = []
        for label in semantic_mutation_labels:
            candidate = json.loads(
                witness._canonical_bytes(semantic_payload).decode("utf-8")
            )
            result_fixture_mutated = False
            if label == "acceptance_true":
                candidate["acceptance_eligible"] = True
            elif label == "artifact_status_drift":
                candidate["artifact_status"] = "PASS"
            elif label == "audit_code_binding":
                candidate["audit_code"]["members"][0]["sha256"] = "0" * 64
            elif label in AUDIT_CODE_SEMANTIC_MUTATION_LABELS:
                role = label.removeprefix("audit_code_role_")
                fabricated = candidate["audit_code"]
                selected = next(
                    member for member in fabricated["members"]
                    if member["role"] == role
                )
                selected["bytes"] += 1
                selected["sha256"] = hashlib.sha256(
                    f"coherent-fabricated-audit-ledger:{role}".encode("utf-8")
                ).hexdigest()
                fabricated["canonical_members_sha256"] = hashlib.sha256(
                    witness._canonical_bytes(fabricated["members"])
                ).hexdigest()
                fabricated_preterminal = json.loads(
                    semantic_result_baseline.decode("utf-8")
                )
                fabricated_preterminal["audit_code"] = json.loads(
                    witness._canonical_bytes(fabricated).decode("utf-8")
                )
                semantic_files[RESULT_NAME].write_bytes(
                    witness._canonical_bytes(fabricated_preterminal, newline=True)
                )
                result_fixture_mutated = True
                fabricated_captures, fabricated_manifest = (
                    _capture_flat_output_universe(
                        semantic_output,
                        sorted(semantic_files),
                        forbidden_physical_aliases=(),
                    )
                )
                candidate["preterminal_result"] = (
                    fabricated_captures[RESULT_NAME].identity()
                )
                candidate["complete_output_artifact_manifest"] = fabricated_manifest
                _validate_terminal_completion_contract(
                    semantic_output / COMPLETION_NAME,
                    candidate,
                    all_true,
                    expected_leg="excel_vs_raw_tsn",
                    expected_audit_code_manifest=fabricated,
                )
            elif label == "audit_drift":
                candidate["audit"] = "fabricated completion audit"
            elif label == "audit_mutation_evidence":
                candidate["audit_code_mutations"][
                    "every_copied_role_byte_mutation_rejected"
                ] = False
            elif label == "containment_evidence":
                candidate["output_containment_mutations"]["same_root_rejected"] = False
            elif label == "expected_final_names":
                candidate["expected_final_artifact_names"].append(
                    "undeclared-final-name.bin"
                )
            elif label == "extra_top_level":
                candidate["unexpected"] = True
            elif label == "family_acceptance_true":
                candidate["stage8_family_accepted"] = True
            elif label == "invariant_false":
                candidate["invariants"][
                    sorted(COMPLETION_INVARIANT_KEYS)[0]
                ] = False
            elif label == "manifest_aggregate":
                candidate["complete_output_artifact_manifest"]["bytes"] += 1
            elif label == "manifest_self_inclusion":
                members = candidate["complete_output_artifact_manifest"]["members"]
                members.append({
                    "relative_path": COMPLETION_NAME,
                    "bytes": 0,
                    "sha256": hashlib.sha256(b"").hexdigest(),
                })
                members.sort(key=lambda item: item["relative_path"])
                candidate["complete_output_artifact_manifest"]["files"] = len(members)
                candidate["complete_output_artifact_manifest"]["bytes"] = sum(
                    item["bytes"] for item in members
                )
                candidate["complete_output_artifact_manifest"][
                    "canonical_members_sha256"
                ] = hashlib.sha256(witness._canonical_bytes(members)).hexdigest()
            elif label == "missing_audit":
                del candidate["audit"]
            elif label == "output_root_drift":
                candidate["output_root"] = str(semantic_output.parent)
            elif label == "payload_chunk_universe":
                candidate["payload_chunk_names"] = []
            elif label == "post_result_phase":
                candidate["post_result_input_revalidation"]["phase"] = (
                    "before nonterminal result publication"
                )
            elif label == "post_result_tree":
                candidate["post_result_input_revalidation"]["trees"] = [
                    {"fabricated_tree": True}
                ]
            elif label == "preterminal_result_identity":
                candidate["preterminal_result"]["sha256"] = "0" * 64
            elif label == "publication_controls_extra_key":
                candidate["publication_lifecycle_mutations"]["unexpected"] = True
            elif label == "publication_truth_false":
                candidate["publication_lifecycle_mutations"][
                    "unchanged_baseline_accepted"
                ] = False
            elif label == "schema_version_drift":
                candidate["schema_version"] = 2
            elif label == "status_drift":
                candidate["status"] = "PASS"
            elif label == "terminal_false":
                candidate["terminal"] = False
            elif label == "terminal_precondition_false":
                candidate["terminal_preconditions"][
                    "complete_output_members_physically_disjoint"
                ] = False
            elif label == "terminal_residue":
                candidate["terminal_residue_gate"]["transient_residue"] = [
                    "unexpected-residue.tmp"
                ]
            elif label == "wrong_valid_leg":
                candidate["leg"] = "pdf_vs_raw_tsn"
                wrong_leg_preterminal = json.loads(
                    semantic_result_baseline.decode("utf-8")
                )
                wrong_leg_preterminal["leg"] = "pdf_vs_raw_tsn"
                semantic_files[RESULT_NAME].write_bytes(
                    witness._canonical_bytes(wrong_leg_preterminal, newline=True)
                )
                result_fixture_mutated = True
                wrong_leg_captures, wrong_leg_manifest = (
                    _capture_flat_output_universe(
                        semantic_output,
                        sorted(semantic_files),
                        forbidden_physical_aliases=(),
                    )
                )
                candidate["preterminal_result"] = (
                    wrong_leg_captures[RESULT_NAME].identity()
                )
                candidate["complete_output_artifact_manifest"] = wrong_leg_manifest
                _validate_terminal_completion_contract(
                    semantic_output / COMPLETION_NAME,
                    candidate,
                    all_true,
                    expected_leg="pdf_vs_raw_tsn",
                    expected_audit_code_manifest=audit_code_manifest,
                )
            else:
                raise DirectRawWitnessError(f"unimplemented completion mutation: {label}")
            rejected = False
            try:
                _publish_terminal_completion(
                    semantic_output / COMPLETION_NAME,
                    candidate,
                    preconditions=all_true,
                    expected_leg="excel_vs_raw_tsn",
                    expected_audit_code_manifest=audit_code_manifest,
                    forbidden_physical_aliases=(),
                    final_validation=lambda _phase: None,
                )
            except DirectRawWitnessError as exc:
                if label in AUDIT_CODE_SEMANTIC_MUTATION_LABELS:
                    _require(
                        str(exc)
                        == (
                            "completion audit-code ledger is not the freshly "
                            "captured physical ledger"
                        ),
                        f"coherent audit-ledger mutation hit the wrong branch: {label}",
                    )
                elif label == "wrong_valid_leg":
                    _require(
                        str(exc)
                        == (
                            "detached completion leg is not the externally "
                            "requested exact leg"
                        ),
                        "coherent wrong-leg mutation hit the wrong validator branch",
                    )
                rejected = True
            finally:
                if result_fixture_mutated:
                    semantic_files[RESULT_NAME].write_bytes(semantic_result_baseline)
            if not rejected:
                raise DirectRawWitnessError(
                    f"completion semantic mutation published terminal PASS: {label}"
                )
            require_no_terminal(semantic_output, f"completion semantic mutation {label}")
            restored_captures, _restored_manifest = _capture_flat_output_universe(
                semantic_output,
                sorted(semantic_files),
                forbidden_physical_aliases=(),
            )
            _require(
                restored_captures[RESULT_NAME].payload == semantic_result_baseline,
                f"completion semantic mutation did not restore result fixture: {label}",
            )
            rejected_completion_semantics.append(label)
        _require(
            set(rejected_completion_semantics) == COMPLETION_SEMANTIC_MUTATION_LABELS,
            "completion semantic mutation census drift",
        )

        protected_zero = base / "protected-zero-byte-input.bin"
        protected_zero.write_bytes(b"")
        hardlink_output, hardlink_files, *_hardlink = prepare(
            "bound-input-hardlink-control"
        )
        hardlink_lease = hardlink_files[witness.PUBLICATION_LEASE_NAME]
        hardlink_lease.unlink()
        try:
            os.link(protected_zero, hardlink_lease)
        except OSError as exc:
            raise DirectRawWitnessError(
                "Windows hardlink-to-bound-input mutation could not execute"
            ) from exc
        bound_input_hardlink_rejected = False
        try:
            _capture_flat_output_universe(
                hardlink_output, sorted(hardlink_files),
                forbidden_physical_aliases=(protected_zero,),
            )
        except DirectRawWitnessError:
            bound_input_hardlink_rejected = True
        _require(
            bound_input_hardlink_rejected,
            "zero-byte output lease hardlinked to a bound input escaped",
        )

        pairwise_output, pairwise_files, *_pairwise = prepare(
            "pairwise-output-hardlink-control"
        )
        pairwise_source = pairwise_files["comparison.xlsx"]
        pairwise_alias = pairwise_files["comparison (values).xlsx"]
        pairwise_alias.unlink()
        try:
            os.link(pairwise_source, pairwise_alias)
        except OSError as exc:
            raise DirectRawWitnessError(
                "Windows pairwise-output hardlink mutation could not execute"
            ) from exc
        pairwise_output_hardlink_rejected = False
        try:
            _capture_flat_output_universe(
                pairwise_output, sorted(pairwise_files),
                forbidden_physical_aliases=(),
            )
        except DirectRawWitnessError:
            pairwise_output_hardlink_rejected = True
        _require(
            pairwise_output_hardlink_rejected,
            "two declared final output names hardlinked together escaped",
        )

        success_output, success_files, *_success_parts, success_payload = prepare(
            "successful-terminal-control"
        )
        callback_phases: list[str] = []
        precommit_stdout = io.StringIO()
        with contextlib.redirect_stdout(precommit_stdout):
            success_rc = _publish_terminal_completion(
                success_output / COMPLETION_NAME,
                success_payload,
                preconditions=all_true,
                expected_leg="excel_vs_raw_tsn",
                expected_audit_code_manifest=audit_code_manifest,
                forbidden_physical_aliases=(),
                final_validation=lambda phase: callback_phases.append(phase),
            )
        precommit_lines = [
            line for line in precommit_stdout.getvalue().splitlines() if line.strip()
        ]
        _require(len(precommit_lines) == 1, "successful control precommit stdout drift")
        precommit_record = json.loads(precommit_lines[0])
        _require(
            precommit_record.get("status")
            == "TERMINAL_COMPLETION_PREPARED_NOT_COMMITTED"
            and precommit_record.get("terminal") is False,
            "successful control emitted terminal/PASS stdout before commit",
        )
        _require(success_rc == 0, "successful terminal publication control returned nonzero")
        success_names = sorted([*success_files, COMPLETION_NAME])
        success_captures, _success_manifest = _capture_flat_output_universe(
            success_output, success_names, forbidden_physical_aliases=(),
        )
        completion_bytes = witness._canonical_bytes(success_payload, newline=True)
        _require(
            success_captures[COMPLETION_NAME].payload == completion_bytes
            and not pending_names(success_output)
            and callback_phases
            == ["after_completion_staging", "immediately_before_terminal_commit"],
            "successful terminal publication control did not commit exactly",
        )

        callback_output, callback_files, *_callback_parts, callback_payload = prepare(
            "second-callback-artifact-mutation-control"
        )
        callback_target = callback_files[RESULT_NAME]

        def mutate_on_second_callback(phase: str) -> None:
            if phase == "immediately_before_terminal_commit":
                callback_target.write_bytes(b"second-callback-mutation")

        second_callback_mutation_rejected = False
        try:
            _publish_terminal_completion(
                callback_output / COMPLETION_NAME,
                callback_payload,
                preconditions=all_true,
                expected_leg="excel_vs_raw_tsn",
                expected_audit_code_manifest=audit_code_manifest,
                forbidden_physical_aliases=(),
                final_validation=mutate_on_second_callback,
            )
        except DirectRawWitnessError:
            require_no_terminal(callback_output, "second-callback artifact mutation")
            second_callback_mutation_rejected = True
        _require(
            second_callback_mutation_rejected,
            "second-callback artifact mutation escaped final recapture",
        )

        staging_output, _staging_files, *_staging_parts, staging_payload = prepare(
            "staging-payload-mutation-control"
        )

        def mutate_staging(phase: str) -> None:
            if phase != "after_completion_staging":
                return
            staged = [
                item for item in staging_output.iterdir()
                if item.name.startswith(f".{COMPLETION_NAME}.")
                and item.name.endswith(".pending")
            ]
            _require(len(staged) == 1, "staging mutation control could not locate candidate")
            with staged[0].open("ab") as handle:
                handle.write(b"staging-mutation")
                handle.flush()
                os.fsync(handle.fileno())

        staging_payload_mutation_rejected = False
        try:
            _publish_terminal_completion(
                staging_output / COMPLETION_NAME,
                staging_payload,
                preconditions=all_true,
                expected_leg="excel_vs_raw_tsn",
                expected_audit_code_manifest=audit_code_manifest,
                forbidden_physical_aliases=(),
                final_validation=mutate_staging,
            )
        except DirectRawWitnessError:
            require_no_terminal(staging_output, "staging payload mutation")
            staging_payload_mutation_rejected = True
        _require(
            staging_payload_mutation_rejected,
            "staging payload mutation escaped identity recapture",
        )

        residue_output, _residue_files, *_residue_parts, residue_payload = prepare(
            "staging-residue-mutation-control"
        )

        def add_residue(phase: str) -> None:
            if phase == "after_completion_staging":
                (residue_output / "unexpected-residue.tmp").write_bytes(b"residue")

        staging_residue_rejected = False
        try:
            _publish_terminal_completion(
                residue_output / COMPLETION_NAME,
                residue_payload,
                preconditions=all_true,
                expected_leg="excel_vs_raw_tsn",
                expected_audit_code_manifest=audit_code_manifest,
                forbidden_physical_aliases=(),
                final_validation=add_residue,
            )
        except DirectRawWitnessError:
            _require(
                not os.path.lexists(residue_output / COMPLETION_NAME),
                "staging residue mutation left a terminal completion",
            )
            _require(not pending_names(residue_output), "staging residue left pending file")
            staging_residue_rejected = True
        _require(staging_residue_rejected, "staging residue escaped exact universe gate")

        collision_output, _collision_files, *_collision_parts, collision_payload = prepare(
            "destination-collision-control"
        )
        collision = collision_output / COMPLETION_NAME
        collision_raw = b"preexisting-destination-must-not-be-overwritten"
        collision.write_bytes(collision_raw)
        destination_collision_rejected = False
        try:
            _publish_terminal_completion(
                collision,
                collision_payload,
                preconditions=all_true,
                expected_leg="excel_vs_raw_tsn",
                expected_audit_code_manifest=audit_code_manifest,
                forbidden_physical_aliases=(),
                final_validation=lambda _phase: None,
            )
        except DirectRawWitnessError:
            _require(
                collision.read_bytes() == collision_raw
                and not pending_names(collision_output),
                "terminal destination collision overwrote or left staging residue",
            )
            destination_collision_rejected = True
        _require(destination_collision_rejected, "existing completion destination was replaced")

        symlink_output = base / "final-artifact-symlink-control"
        symlink_output.mkdir()
        artifact = symlink_output / "artifact.bin"
        symlink_status = "unexecuted_platform_denied"
        target = base / "outside-target.bin"
        target.write_bytes(b"outside")
        try:
            os.symlink(target, artifact, target_is_directory=False)
        except (OSError, NotImplementedError):
            artifact.write_bytes(b"restored")
        else:
            try:
                _capture_flat_output_universe(
                    symlink_output, [artifact.name], forbidden_physical_aliases=(),
                )
            except DirectRawWitnessError:
                symlink_status = "executed_and_rejected"
            else:
                raise DirectRawWitnessError(
                    "final-artifact symlink mutation escaped lexical lstat gate"
                )

        rejected_preconditions: list[str] = []
        for failed in sorted(TERMINAL_PRECONDITION_KEYS):
            preconditions = {key: True for key in TERMINAL_PRECONDITION_KEYS}
            preconditions[failed] = False
            precondition_output = base / f"failed-precondition-{failed}"
            precondition_output.mkdir()
            completion = precondition_output / COMPLETION_NAME
            try:
                _publish_terminal_completion(
                    completion, {"status": "MUST_NOT_PUBLISH"},
                    preconditions=preconditions,
                    expected_leg="excel_vs_raw_tsn",
                    expected_audit_code_manifest=audit_code_manifest,
                    forbidden_physical_aliases=(),
                    final_validation=lambda _phase: None,
                )
            except DirectRawWitnessError:
                _require(
                    not os.path.lexists(completion),
                    f"failed terminal precondition left a completion: {failed}",
                )
                rejected_preconditions.append(failed)
            else:
                raise DirectRawWitnessError(
                    f"failed terminal precondition published PASS: {failed}"
                )
        successful_control = {
            "exit_zero": success_rc == 0,
            "precommit_stdout_nonterminal": True,
            "validation_phases": callback_phases,
            "exact_final_universe": sorted(success_captures) == success_names,
            "no_pending_residue": not pending_names(success_output),
            "completion_canonical_and_exact":
                success_captures[COMPLETION_NAME].payload == completion_bytes,
        }
        semantic_census_exact = (
            set(rejected_completion_semantics) == COMPLETION_SEMANTIC_MUTATION_LABELS
        )
        failed_preconditions_exact = (
            set(rejected_preconditions) == TERMINAL_PRECONDITION_KEYS
        )
        all_controls_passed = all([
            unchanged_baseline_accepted,
            completion_contract_baseline_accepted,
            set(rejected_artifact_roles) == set(role_payloads),
            missing_name_rejected,
            extra_name_rejected,
            bound_input_hardlink_rejected,
            pairwise_output_hardlink_rejected,
            all(value is True for key, value in successful_control.items()
                if key != "validation_phases"),
            callback_phases
            == ["after_completion_staging", "immediately_before_terminal_commit"],
            second_callback_mutation_rejected,
            staging_payload_mutation_rejected,
            staging_residue_rejected,
            destination_collision_rejected,
            symlink_status == "executed_and_rejected",
            failed_preconditions_exact,
            semantic_census_exact,
        ])
        _require(all_controls_passed, "publication mutation-control aggregate failed")
        publication_result = {
            "unchanged_baseline_accepted": unchanged_baseline_accepted,
            "completion_contract_baseline_accepted":
                completion_contract_baseline_accepted,
            "post_manifest_artifact_roles_rejected": rejected_artifact_roles,
            "missing_name_universe_rejected": missing_name_rejected,
            "extra_name_universe_rejected": extra_name_rejected,
            "bound_input_hardlink_rejected": bound_input_hardlink_rejected,
            "pairwise_output_hardlink_rejected": pairwise_output_hardlink_rejected,
            "completion_semantic_mutations_rejected":
                rejected_completion_semantics,
            "completion_semantic_mutations_left_no_terminal_file":
                semantic_census_exact,
            "successful_terminal_control": successful_control,
            "second_callback_artifact_mutation_rejected":
                second_callback_mutation_rejected,
            "staging_payload_mutation_rejected": staging_payload_mutation_rejected,
            "staging_residue_rejected": staging_residue_rejected,
            "destination_collision_rejected_without_overwrite":
                destination_collision_rejected,
            "final_artifact_symlink_mutation": symlink_status,
            "failed_terminal_preconditions": rejected_preconditions,
            "all_failed_preconditions_left_no_terminal_file":
                failed_preconditions_exact,
            "disposable_control_only": False,
            "all_publication_controls_passed": all_controls_passed,
        }
        _require_keys(
            publication_result, PUBLICATION_MUTATION_KEYS,
            "derived publication mutation result",
        )
        return publication_result


def _assert_created_output_disjoint(root: Path, preflight: Preflight) -> Path:
    violations = _output_policy_violations(
        root,
        protected_files=preflight.bound_files,
        protected_roots=preflight.protected_roots,
        permit_existing_output_root=True,
    )
    _require(not violations, f"created output root lost containment: {violations}")
    _assert_plain_components(root)
    try:
        facts = os.lstat(root)
    except OSError as exc:
        raise DirectRawWitnessError(f"created output root vanished: {root}") from exc
    _require(stat.S_ISDIR(facts.st_mode), f"created output root is not a directory: {root}")
    return root.resolve(strict=True)


def _create_output_root(candidate: Path, preflight: Preflight) -> Path:
    canonical = _validate_output_root(
        candidate,
        protected_files=preflight.bound_files,
        protected_roots=preflight.protected_roots,
    )
    try:
        candidate.mkdir(parents=False, exist_ok=False)
    except (FileExistsError, OSError) as exc:
        raise DirectRawWitnessError(f"could not create exclusive output root: {candidate}") from exc
    root = _assert_created_output_disjoint(canonical, preflight)
    _require(not any(root.iterdir()), "new output root was not empty")
    return root


def _load_product(
    leg: str,
    raw_tsn_workbook: Path,
) -> tuple[Callable[..., object], Path, Path, object, object]:
    sys.path.insert(0, str(SCRIPTS_ROOT))
    from events import Events  # type: ignore
    import consolidation_meta  # type: ignore
    import compare_highway_sequence_tsn as hsl  # type: ignore

    if leg == "excel_vs_raw_tsn":
        compare = hsl.compare
        side_a = EXCEL_INPUT
    elif leg == "pdf_vs_raw_tsn":
        import compare_highway_sequence_pdf as hsl_pdf  # type: ignore
        compare = hsl_pdf.TSMIS_PDF_VS_TSN.compare
        side_a = PDF_INPUT
    else:
        raise DirectRawWitnessError(f"unsupported direct raw-TSN leg: {leg}")
    return compare, side_a, raw_tsn_workbook, Events(), consolidation_meta


def _write_canonical(path: Path, payload: object) -> dict[str, object]:
    expected = witness._canonical_bytes(payload, newline=True)
    identity = witness._write_exclusive(path, payload)
    _require(
        identity["bytes"] == len(expected)
        and identity["sha256"] == hashlib.sha256(expected).hexdigest()
        and path.read_bytes() == expected,
        f"canonical audit serialization drift: {path}",
    )
    return identity


def _validate_chunk_sets(
    decoded_payload: Mapping[str, object],
    residue: Mapping[str, object],
) -> set[str]:
    chunks = decoded_payload.get("chunks")
    _require(isinstance(chunks, list) and chunks, "decoded payload chunk ledger absent")
    decoded = {str(item["relative_path"]) for item in chunks}
    inventoried = set(
        residue["exact_artifact_universe"]["payload_chunks"]
    )
    _require(decoded == inventoried, "referenced/decoded/inventoried payload chunk sets disagree")
    return decoded


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run one non-accepting Highway Sequence product witness against an "
            "exact direct-source 69,804-row raw-TSN twin."
        )
    )
    parser.add_argument(
        "--leg", choices=LEG_CHOICES, required=True, default=None,
        action=witness._SingleValue,
    )
    parser.add_argument(
        "--twin-root", type=Path, required=True, default=None,
        action=witness._SingleValue,
    )
    parser.add_argument(
        "--output-root", type=Path, required=True, default=None,
        action=witness._SingleValue,
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    _require(
        importlib.metadata.version("openpyxl") == "3.1.5"
        and importlib.metadata.version("pdfplumber") == "0.11.9",
        "direct raw product runner audit-runtime drift",
    )
    _assert_product_not_loaded()
    audit_code_before = _capture_audit_code()
    audit_code_manifest = _audit_code_manifest(audit_code_before)
    audit_code_mutations = _run_audit_code_mutation_probe(audit_code_before)
    containment_mutations = _run_disjointness_mutation_probes()
    publication_mutations = _run_publication_mutation_probes(
        audit_code_manifest=audit_code_manifest,
        audit_code_mutations=audit_code_mutations,
        containment_mutations=containment_mutations,
    )
    preflight = _validate_direct_twin(args.twin_root)
    forbidden_output_aliases = tuple(sorted(
        {
            *preflight.bound_files,
            *(captured.path for captured in audit_code_before.values()),
        },
        key=lambda path: str(path).casefold(),
    ))
    _assert_product_not_loaded()
    lease_policy = witness._source_backed_lease_policy()
    root = _create_output_root(args.output_root, preflight)
    after_output_creation = _revalidate_frozen_inputs(
        preflight, phase="after exclusive output-root creation",
    )
    _assert_created_output_disjoint(root, preflight)

    compare, side_a, side_b, events, consolidation_meta = _load_product(
        args.leg, preflight.raw_tsn_workbook,
    )
    outputs = witness._comparison_paths(root)
    result = compare(
        side_a,
        side_b,
        outputs["formulas"],
        events=events,
        confirm_overwrite=lambda _path: False,
        mode="both",
    )
    result_summary, declared_outputs = witness._validate_product_result(
        result, outputs, consolidation_meta,
    )
    decoded_payload = dev._decode_payload_chunks(
        outputs, result, consolidation_meta,
    )
    residue = witness._residue_gate(
        root, lease_policy, allowed_audit_names=set(),
    )
    chunk_names = _validate_chunk_sets(decoded_payload, residue)
    after_product_publication = _revalidate_frozen_inputs(
        preflight, phase="after product publication",
    )
    _assert_created_output_disjoint(root, preflight)

    normalized_leg = (
        "excel_vs_normalized_tsn"
        if args.leg == "excel_vs_raw_tsn"
        else "pdf_vs_normalized_tsn"
    )
    loaded_product_code = witness._loaded_product_manifest(normalized_leg)
    _assert_created_output_disjoint(root, preflight)
    product_manifest_identity = _write_canonical(
        root / PRODUCT_CODE_MANIFEST_NAME, loaded_product_code,
    )
    residue = witness._residue_gate(
        root,
        lease_policy,
        allowed_audit_names={PRODUCT_CODE_MANIFEST_NAME},
    )
    _require(
        _validate_chunk_sets(decoded_payload, residue) == chunk_names,
        "payload chunk universe changed after product-code manifest",
    )

    artifact_manifest_before_result = witness._artifact_manifest(
        root, excluded_names={ARTIFACT_MANIFEST_NAME, RESULT_NAME},
    )
    _assert_created_output_disjoint(root, preflight)
    artifact_manifest_identity = _write_canonical(
        root / ARTIFACT_MANIFEST_NAME, artifact_manifest_before_result,
    )
    residue = witness._residue_gate(
        root,
        lease_policy,
        allowed_audit_names={
            PRODUCT_CODE_MANIFEST_NAME, ARTIFACT_MANIFEST_NAME,
        },
    )
    _require(
        _validate_chunk_sets(decoded_payload, residue) == chunk_names,
        "payload chunk universe changed after artifact manifest",
    )
    before_result = _revalidate_frozen_inputs(
        preflight, phase="immediately before terminal result publication",
    )
    _assert_created_output_disjoint(root, preflight)

    expected_precompletion_names = sorted(
        set(residue["exact_artifact_universe"]["core_names"])
        | set(chunk_names)
        | {
            PRODUCT_CODE_MANIFEST_NAME,
            ARTIFACT_MANIFEST_NAME,
            RESULT_NAME,
        }
    )
    payload = {
        "schema_version": 1,
        "audit": PRETERMINAL_AUDIT,
        "status": PRETERMINAL_STATUS,
        "terminal": False,
        "artifact_status": ARTIFACT_STATUS,
        "acceptance_eligible": False,
        "stage8_family_accepted": False,
        "reason": (
            "Post-result input-tree, full-artifact, and containment checks must pass "
            "before a separate detached terminal completion may be committed."
        ),
        "leg": args.leg,
        "output_root": str(root),
        "direct_twin_preimport_validation": preflight.summary,
        "audit_code": audit_code_manifest,
        "audit_code_mutations": audit_code_mutations,
        "output_containment_mutations": containment_mutations,
        "publication_lifecycle_mutations": publication_mutations,
        "input_tree_revalidation": {
            "after_output_creation": after_output_creation,
            "after_product_publication": after_product_publication,
            "before_result_publication": before_result,
            "post_result_revalidation_not_yet_performed": True,
        },
        "result": result_summary,
        "outputs": declared_outputs["workbooks"],
        "outcome_sidecars": declared_outputs["outcome_sidecars"],
        "decoded_comparison_payload": decoded_payload,
        "publication_artifacts": {
            "payload_chunks": decoded_payload["chunks"],
            "outcome_sidecars": declared_outputs["outcome_sidecars"],
            "permanent_lease": residue["permanent_lease"],
        },
        "residue_gate": residue,
        "loaded_product_code": loaded_product_code,
        "product_code_manifest": product_manifest_identity,
        "artifact_manifest_before_result": artifact_manifest_before_result,
        "artifact_manifest": artifact_manifest_identity,
        "deterministic_serialization": {
            "json": "UTF-8 canonical JSON; sorted keys; compact separators; LF terminator",
            "product_code_manifest_canonical": True,
            "artifact_manifest_canonical": True,
            "result_canonical": True,
        },
        "expected_precompletion_artifact_names": expected_precompletion_names,
        "required_detached_terminal_completion": COMPLETION_NAME,
        "invariants": {
            "one_leg_only": args.leg in LEG_CHOICES,
            "direct_twin_v1_validated_before_product_import": True,
            "direct_twin_current_builder_exact": True,
            "direct_twin_not_family_acceptance": True,
            "raw_records_69804":
                preflight.summary["counts"]["raw_records"] == 69_804,
            "tsmis_inputs_independently_exact_bound": True,
            "accepted_stage6_and_raw_pdf_bindings_exact": True,
            "bidirectional_998_998_zero_orphan_topology": True,
            "reverse_only_topology_mutation_rejected": True,
            "workbook_reopen_equals_all_provenance_rows": True,
            "two_way_output_input_disjointness": True,
            "disposable_containment_mutations_passed": True,
            "publication_lifecycle_mutations_passed":
                publication_mutations[
                    "all_publication_controls_passed"
                ] is True,
            "input_tree_universes_frozen_through_pre_result": True,
            "complete_ok_zero_zero": True,
            "pairing_exact": True,
            "committed_formula_value_twin": True,
            "two_trusted_outcome_sidecars": True,
            "payload_chunks_decoded_and_bound": True,
            "referenced_decoded_inventoried_chunks_equal": True,
            "exact_artifact_universe_declared": True,
            "no_transient_residue": not residue["transient_residue"],
            "only_zero_byte_source_backed_permanent_lease":
                residue["permanent_lease"]["bytes"] == 0,
            "canonical_deterministic_audit_json": True,
            "loaded_product_code_manifested": True,
            "no_delete_or_overwrite": True,
            "this_record_is_explicitly_nonterminal": True,
        },
    }
    _require(
        all(payload["invariants"].values()),
        f"direct raw product witness invariants failed: {payload['invariants']}",
    )
    result_identity = _write_canonical(root / RESULT_NAME, payload)

    final_residue = witness._residue_gate(
        root,
        lease_policy,
        allowed_audit_names={
            PRODUCT_CODE_MANIFEST_NAME, ARTIFACT_MANIFEST_NAME, RESULT_NAME,
        },
    )
    final_names = sorted(path.name for path in root.iterdir())
    _require(
        final_names == expected_precompletion_names,
        "precompletion output artifact universe drift",
    )
    _require(
        set(final_residue["exact_artifact_universe"]["payload_chunks"])
        == chunk_names,
        "precompletion payload chunk universe drift",
    )
    post_result = _revalidate_frozen_inputs(
        preflight, phase="after nonterminal result publication",
    )
    _assert_created_output_disjoint(root, preflight)

    expected_identities: dict[str, tuple[int, str]] = {}

    def add_declared(identity: Mapping[str, object], label: str) -> None:
        name, size, digest = _declared_identity(
            identity, root=root, label=label,
        )
        observed = expected_identities.get(name)
        _require(
            observed is None or observed == (size, digest),
            f"conflicting declared identities for final artifact: {name}",
        )
        expected_identities[name] = (size, digest)

    for flavor, identity in declared_outputs["workbooks"].items():
        add_declared(identity, f"{flavor} workbook")
    for flavor, identity in declared_outputs["outcome_sidecars"].items():
        add_declared(identity, f"{flavor} outcome sidecar")
    add_declared(final_residue["permanent_lease"], "permanent publication lease")
    add_declared(product_manifest_identity, "product-code manifest")
    add_declared(artifact_manifest_identity, "artifact manifest")
    add_declared(result_identity, "nonterminal result")
    for descriptor in decoded_payload["chunks"]:
        name = str(descriptor["relative_path"])
        _require(
            Path(name).name == name and name in chunk_names,
            f"unsafe final payload-chunk declaration: {name!r}",
        )
        identity = (
            int(descriptor["compressed_bytes"]),
            str(descriptor["compressed_sha256"]),
        )
        _require(
            name not in expected_identities
            or expected_identities[name] == identity,
            f"conflicting final payload-chunk identity: {name}",
        )
        expected_identities[name] = identity

    complete_captures, complete_output_manifest = _capture_flat_output_universe(
        root, expected_precompletion_names,
        forbidden_physical_aliases=forbidden_output_aliases,
    )
    _validate_complete_output_identities(
        complete_captures, expected_identities,
    )
    _require(
        _strict_json_bytes(
            complete_captures[PRODUCT_CODE_MANIFEST_NAME].payload,
            PRODUCT_CODE_MANIFEST_NAME,
            canonical=True,
        ) == loaded_product_code,
        "final product-code manifest semantic drift",
    )
    _require(
        _strict_json_bytes(
            complete_captures[ARTIFACT_MANIFEST_NAME].payload,
            ARTIFACT_MANIFEST_NAME,
            canonical=True,
        ) == artifact_manifest_before_result,
        "final artifact manifest semantic drift",
    )
    _require(
        _strict_json_bytes(
            complete_captures[RESULT_NAME].payload,
            RESULT_NAME,
            canonical=True,
        ) == payload,
        "final nonterminal result semantic drift",
    )
    audit_code_before_completion = _revalidate_audit_code(audit_code_before)
    _require(
        audit_code_before_completion == audit_code_manifest,
        "audit-code manifest drift before completion staging",
    )

    terminal_preconditions = {
        "post_result_input_trees_exact": True,
        "complete_output_artifact_universe_exact":
            set(complete_captures) == set(expected_precompletion_names),
        "complete_output_identities_exact":
            set(complete_captures) == set(expected_identities),
        "complete_output_members_physically_disjoint": True,
        "terminal_residue_exact": not final_residue["transient_residue"],
        "output_root_still_disjoint": True,
    }
    _require_terminal_preconditions(terminal_preconditions)
    expected_final_names = sorted([
        *expected_precompletion_names, COMPLETION_NAME,
    ])
    completion_payload = {
        "schema_version": 1,
        "audit": COMPLETION_AUDIT,
        "status": COMPLETION_STATUS,
        "terminal": True,
        "artifact_status": ARTIFACT_STATUS,
        "acceptance_eligible": False,
        "stage8_family_accepted": False,
        "leg": args.leg,
        "output_root": str(root),
        "preterminal_result": result_identity,
        "audit_code": audit_code_manifest,
        "audit_code_mutations": audit_code_mutations,
        "post_result_input_revalidation": post_result,
        "complete_output_artifact_manifest": complete_output_manifest,
        "terminal_residue_gate": final_residue,
        "payload_chunk_names": sorted(chunk_names),
        "terminal_preconditions": terminal_preconditions,
        "output_containment_mutations": containment_mutations,
        "publication_lifecycle_mutations": publication_mutations,
        "expected_final_artifact_names": expected_final_names,
        "invariants": {
            "nonterminal_result_exactly_bound": True,
            "post_result_input_trees_exact": True,
            "complete_output_artifact_universe_exact": True,
            "complete_output_identities_exact": True,
            "complete_output_members_physically_disjoint": True,
            "referenced_decoded_inventoried_final_chunks_equal": True,
            "terminal_residue_exact": True,
            "output_root_still_disjoint": True,
            "terminal_completion_is_last_commit": True,
            "not_family_acceptance": True,
        },
    }
    _require(all(completion_payload["invariants"].values()), "completion invariant drift")
    def final_validation(phase: str) -> None:
        _revalidate_frozen_inputs(
            preflight, phase=f"terminal staging: {phase}",
        )
        _assert_created_output_disjoint(root, preflight)
        observed_code = _revalidate_audit_code(audit_code_before)
        _require(
            observed_code == audit_code_manifest,
            f"audit-code drift during terminal staging: {phase}",
        )

    return _publish_terminal_completion(
        root / COMPLETION_NAME,
        completion_payload,
        preconditions=terminal_preconditions,
        expected_leg=args.leg,
        expected_audit_code_manifest=audit_code_before_completion,
        forbidden_physical_aliases=forbidden_output_aliases,
        final_validation=final_validation,
    )


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except witness.WitnessError as exc:
        print(f"FAIL Highway Sequence direct raw-twin product leg: {exc}")
        raise SystemExit(1)
