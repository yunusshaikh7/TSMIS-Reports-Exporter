"""Phase-3 E1 red/green gate for the canonical compared-cell policy.

This check deliberately covers BOTH halves of the contract:

* ``normalize_value`` + public ``compared_cell`` are the Python authority;
* the formulas workbook must mirror that authority without Excel coercion,
  display-marker inference, or precision loss.

The policy is intentionally narrow.  Text is case-sensitive; only actual Python
booleans fold to ``TRUE``/``FALSE``; TRIM means ASCII U+0020 only; Med-Wid accepts
only an ASCII numeric core and an optional printable-ASCII non-digit/non-dot suffix;
blank is not zero; error-looking strings and the visible difference separator are data;
all finite source numerics are emitted as exact comparison text so Excel cannot
rewrite scale/exponent/precision, while non-finite numerics fail closed.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_equality_policy.py

Optional installed-Excel evaluation (COM; default remains hermetic):
    build\\.venv\\Scripts\\python.exe build\\check_compare_equality_policy.py --excel
"""
from __future__ import annotations

import argparse
from decimal import Decimal
from pathlib import Path

from _checklib import Checker, patch, scripts_path, temp_dir

scripts_path()

import compare_core as cc
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ERROR_CODES


MARK = " ≠ "

TEXT_SCHEMA = cc.CompareSchema(
    report_name="Equality Policy",
    header=["Loc", "Text"],
    side_a="LEFT",
    side_b="RIGHT",
    id_noun="row",
    id_noun_plural="rows",
)
MED_SCHEMA = cc.CompareSchema(
    report_name="Equality Policy Med-Wid",
    header=["Loc", "Med-Wid"],
    side_a="LEFT",
    side_b="RIGHT",
    id_noun="row",
    id_noun_plural="rows",
    medwid_fields=("Med-Wid",),
)
DITTO_SCHEMA = cc.CompareSchema(
    report_name="Equality Policy Ditto",
    header=["Loc", "Value"],
    side_a="LEFT",
    side_b="RIGHT",
    id_noun="row",
    id_noun_plural="rows",
    ditto_nonasserting=True,
)
WORKBOOK_SCHEMA = cc.CompareSchema(
    report_name="Equality Policy Workbook",
    header=["Loc", "Text", "Med-Wid"],
    side_a="LEFT",
    side_b="RIGHT",
    id_noun="row",
    id_noun_plural="rows",
    medwid_fields=("Med-Wid",),
)


class IntOne(int):
    """An int subclass must not fall into an ``isinstance(v, int)`` bool fold."""


def _cell_equality(result):
    """Read current tuple or the planned typed compared-cell result.

    E1 may replace the compatibility tuple with a typed value.  The semantic
    check should survive that mechanical migration while still requiring an
    explicit equality field on the new type.
    """
    if isinstance(result, tuple):
        if len(result) != 3:
            raise TypeError(f"unexpected compared_cell tuple: {result!r}")
        return result[2]
    for name in ("equal", "is_equal", "equality"):
        if hasattr(result, name):
            return getattr(result, name)
    raise TypeError(f"compared_cell result has no equality field: {result!r}")


def _equal(schema, left, right):
    left = cc.normalize_value(left)
    right = cc.normalize_value(right)
    result = cc.compared_cell(schema, 1, ["K", left], ["K", right], 0)
    return _cell_equality(result), left, right, result


def _asserting(result):
    if hasattr(result, "asserting"):
        return result.asserting
    return _cell_equality(result) is not None


def _expect(c, name, schema, left, right, want):
    try:
        got, nl, nr, result = _equal(schema, left, right)
        c.check(name, got is want,
                f"left={left!r}->{nl!r}; right={right!r}->{nr!r}; "
                f"equality={got!r}; result={result!r}")
    except Exception as exc:  # one missing seam must not hide the other cases
        c.check(name, False, f"policy raised {type(exc).__name__}: {exc}")


def _header_col(ws, label):
    for cell in ws[1]:
        if cell.value == label:
            return cell.column
    raise KeyError(f"{label!r} is not a header in {ws.title!r}")


def _row_for(ws, key):
    col = _header_col(ws, "Loc")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, col).value == key:
            return row
    raise KeyError(f"Loc={key!r} not found in {ws.title!r}")


def _cell_for(wb, sheet, key, field):
    ws = wb[sheet]
    return ws.cell(_row_for(ws, key), _header_col(ws, field))


def _formula_cells(ws):
    return [str(cell.value) for row in ws.iter_rows() for cell in row
            if cell.data_type == "f" or
            (isinstance(cell.value, str) and cell.value.startswith("="))]


def _cf_formulas(ws):
    formulas = []
    for region in ws.conditional_formatting:
        for rule in ws.conditional_formatting[region]:
            formulas.extend(str(value) for value in (rule.formula or ()))
    return formulas


def _has_marker_state_scan(formula):
    compact = str(formula).replace(" ", "").upper()
    marker = MARK.replace(" ", "")
    return (("SEARCH(" in compact or "FIND(" in compact or "COUNTIF(" in compact)
            and marker in compact)


def test_python_policy(c):
    print("Python compared-cell policy:")

    # Case is data.  Excel '=' is case-insensitive, so the formulas half is
    # checked separately for EXACT().
    _expect(c, "text equality is case-sensitive", TEXT_SCHEMA,
            "ABC", "abc", False)

    # Boolean normalization belongs at the public load seam and MUST precede
    # generic numeric handling: bool is a subclass of int in Python.
    c.check("actual True normalizes to uppercase text TRUE",
            cc.normalize_value(True) == "TRUE",
            repr(cc.normalize_value(True)))
    c.check("actual False normalizes to uppercase text FALSE",
            cc.normalize_value(False) == "FALSE",
            repr(cc.normalize_value(False)))
    c.check("integer 1 is not folded to TRUE",
            type(cc.normalize_value(1)) is int and cc.normalize_value(1) == 1,
            f"1->{cc.normalize_value(1)!r}; True->{cc.normalize_value(True)!r}")
    c.check("integer 0 is not folded to FALSE",
            type(cc.normalize_value(0)) is int and cc.normalize_value(0) == 0,
            f"0->{cc.normalize_value(0)!r}; False->{cc.normalize_value(False)!r}")
    c.check("an int subclass is not treated as a Boolean",
            cc.normalize_value(IntOne(1)) != "TRUE",
            repr(cc.normalize_value(IntOne(1))))
    _expect(c, "actual True equals exact text TRUE", TEXT_SCHEMA,
            True, "TRUE", True)
    _expect(c, "actual False equals exact text FALSE", TEXT_SCHEMA,
            False, "FALSE", True)
    _expect(c, "numeric 1 does not equal TRUE", TEXT_SCHEMA,
            1, "TRUE", False)
    _expect(c, "numeric 0 does not equal FALSE", TEXT_SCHEMA,
            0, "FALSE", False)
    _expect(c, "lowercase Boolean-looking text remains case-sensitive",
            TEXT_SCHEMA, "true", "TRUE", False)

    # The approved whitespace policy is exactly Excel TRIM's ASCII-space
    # behavior.  It does not globally rewrite control whitespace or NBSP.
    _expect(c, "edge ASCII spaces trim and internal ASCII runs collapse",
            TEXT_SCHEMA, "  ALPHA   BETA  ", "ALPHA BETA", True)
    for label, char in (("tab", "\t"), ("CR", "\r"), ("LF", "\n"),
                        ("NBSP", "\u00a0")):
        _expect(c, f"{label} is not globally folded to ASCII space",
                TEXT_SCHEMA, f"ALPHA{char}BETA", "ALPHA BETA", False)
        _expect(c, f"edge {label} is not stripped as ASCII whitespace",
                TEXT_SCHEMA, f"{char}ALPHA{char}", "ALPHA", False)

    ascii_ditto = cc.compared_cell(
        DITTO_SCHEMA, 1, ["K", "  +  "], ["K", "VALUE"], 0)
    c.check("ASCII-space-wrapped plus remains a nonasserting ditto",
            not _asserting(ascii_ditto), repr(ascii_ditto))
    for label, char in (("tab", "\t"), ("NBSP", "\u00a0")):
        unusual_ditto = cc.compared_cell(
            DITTO_SCHEMA, 1, ["K", f"{char}+{char}"], ["K", "VALUE"], 0)
        c.check(f"{label}-wrapped plus is data, not a ditto marker",
                _asserting(unusual_ditto), repr(unusual_ditto))

    # Blank and numeric zero have distinct typed meanings.
    _expect(c, "blank does not equal numeric zero", TEXT_SCHEMA,
            None, 0, False)
    _expect(c, "blank does not equal text zero", TEXT_SCHEMA,
            None, "0", False)

    # Error-looking source strings remain ordinary literal text.
    for token in ERROR_CODES:
        _expect(c, f"literal Excel error {token} equals itself",
                TEXT_SCHEMA, token, token, True)
    _expect(c, "different literal Excel errors remain different", TEXT_SCHEMA,
            "#N/A", "#VALUE!", False)

    # The visible separator is content when it came from a source cell.
    literal_marker = f"ALPHA{MARK}BETA"
    _expect(c, "literal difference-marker text can be equal source content",
            TEXT_SCHEMA, literal_marker, literal_marker, True)
    _expect(c, "literal marker content still compares normally when changed",
            TEXT_SCHEMA, literal_marker, f"ALPHA{MARK}GAMMA", False)

    # Narrow, decimal-safe Med-Wid grammar.  No float/Excel VALUE round-trip is
    # allowed: it loses adjacent large decimals and accepts too much syntax.
    med_cases = (
        ("unsigned zero padding", "0006V", "6V", True),
        ("unsigned decimal zero padding", "0006.5000V", "6.5V", True),
        ("plain unsigned decimal", "000.5000", "0.5", True),
        ("one punctuation suffix is preserved", "0006#", "6#", True),
        ("suffix case is significant", "06v", "6V", False),
        ("signed values stay raw", "-06V", "-6V", False),
        ("explicit plus stays raw", "+06V", "6V", False),
        ("leading-decimal values stay raw", ".50V", "0.5V", False),
        ("exponent notation stays raw", "6e0V", "6V", False),
        ("multiple suffix characters stay raw", "06##", "6##", False),
        ("non-ASCII digits stay raw", "\u0660\u0666V", "6V", False),
        ("mixed Unicode digit suffix stays raw", "06\u0661", "6\u0661", False),
        ("non-ASCII letter suffix stays raw", "06\u00e9", "6\u00e9", False),
        ("control-character suffix stays raw", "06\t", "6\t", False),
        ("large adjacent integers remain distinct",
         "9007199254740992V", "9007199254740993V", False),
        ("large exact decimal trailing zeros normalize",
         "12345678901234567890.0100V", "12345678901234567890.01V", True),
        ("large adjacent decimals remain distinct",
         "12345678901234567890.010V", "12345678901234567890.011V", False),
    )
    for label, left, right, want in med_cases:
        _expect(c, f"Med-Wid: {label}", MED_SCHEMA, left, right, want)

    # Ordinary compared-cell precision is exact in Python; the workbook writer
    # assertions below ensure Excel receives the same exact text.
    big = 12345678901234567
    _expect(c, ">15-digit integer equals its exact text spelling", TEXT_SCHEMA,
            big, str(big), True)
    _expect(c, "adjacent >15-digit integers remain different", TEXT_SCHEMA,
            big, big + 1, False)


def _workbook_rows():
    marker = f"ALPHA{MARK}BETA"
    big = 12345678901234567
    boundary = 123456789012345
    rows = [
        ("CASE", "ABC", "06V", "abc", "6V"),
        ("BOOL", True, "06V", "TRUE", "6V"),
        ("INT_TRUE", 1, "06V", "TRUE", "6V"),
        ("ZERO_FALSE", 0, "06V", "FALSE", "6V"),
        ("SPACE", "  ALPHA   BETA  ", "0006.5000V", "ALPHA BETA", "6.5V"),
        ("TAB", "ALPHA\tBETA", "06V", "ALPHA BETA", "6V"),
        ("CR", "ALPHA\rBETA", "06V", "ALPHA BETA", "6V"),
        ("LF", "ALPHA\nBETA", "06V", "ALPHA BETA", "6V"),
        ("NBSP", "ALPHA\u00a0BETA", "06V", "ALPHA BETA", "6V"),
        ("BLANK", None, "06V", 0, "6V"),
        ("MARK", marker, "06V", marker, "6V"),
        ("BIGEQ", big, "06V", str(big), "6V"),
        ("BIGDIFF", big, "06V", big + 1, "6V"),
        ("MED_PRECISE", "same", "9007199254740992V",
         "same", "9007199254740993V"),
        # Typed blank must remain distinct from numeric zero through INDEX,
        # Med-Wid staging, the hidden state mask, and Spot Check.
        ("MED_BLANK_ZERO", "same", None, "same", 0),
        ("BOUNDARY", boundary, "06V", str(boundary), "6V"),
        ("DEC_SCALE", Decimal("1.2300"), "06V", "1.2300", "6V"),
        ("DEC_ONE", Decimal("1.0"), "06V", "1.0", "6V"),
        ("DEC_EXP", Decimal("1E+3"), "06V", "1E+3", "6V"),
        ("FLOAT_SMALL", 1e-7, "06V", "1e-07", "6V"),
        ("FLOAT_LARGE", 1e20, "06V", "100000000000000000000", "6V"),
    ]
    rows.extend((f"ERR_{i}", token, "06V", token, "6V")
                for i, token in enumerate(ERROR_CODES, start=1))

    # This is the real loader boundary: dates/booleans are canonicalized before
    # the generic engine receives the rows. Numeric objects stay typed here so
    # the workbook writer's exact-text seam is exercised.
    left = [[key, cc.normalize_value(a), cc.normalize_value(ma)]
            for key, a, ma, _b, _mb in rows]
    right = [[key, cc.normalize_value(b), cc.normalize_value(mb)]
             for key, _a, _ma, b, mb in rows]
    # Exercise the row-level U state in both directions. These are deliberately
    # appended after the matched policy corpus so CASE remains Spot's default.
    left.append(["LEFT_ONLY", "left value", None])
    right.append(["RIGHT_ONLY", "right value", 0])
    return left, right


def _expected_state_masks(rows_left, rows_right):
    """Independent, named truth table for the workbook fixture."""
    keys = {row[0] for row in rows_left} | {row[0] for row in rows_right}
    expected = {key: "EE" for key in keys}
    for key in ("CASE", "INT_TRUE", "ZERO_FALSE", "TAB", "CR", "LF",
                "NBSP", "BLANK", "BIGDIFF"):
        expected[key] = "D" + expected[key][1]
    for key in ("MED_PRECISE", "MED_BLANK_ZERO"):
        expected[key] = expected[key][0] + "D"
    expected["LEFT_ONLY"] = "UU"
    expected["RIGHT_ONLY"] = "UU"
    return expected


def _state_columns(ws):
    return [cell.column for cell in ws[1]
            if str(cell.value).startswith("__CMP_E1_STATE_V1_C")]


def test_workbook_policy(c, run_excel=False):
    print("\nTyped outcome + workbook/formula parity:")
    rows_left, rows_right = _workbook_rows()
    expected_masks = _expected_state_masks(rows_left, rows_right)

    with temp_dir("tsmis_equality_policy_") as tmp:
        values_path = Path(tmp) / "policy-values.xlsx"
        formulas_path = Path(tmp) / "policy-formulas.xlsx"
        rv = cc.run_compare(WORKBOOK_SCHEMA, rows_left, rows_right, False,
                            values_path, mode="values")
        rf = cc.run_compare(WORKBOOK_SCHEMA, rows_left, rows_right, False,
                            formulas_path, mode="formulas")
        c.check("values workbook builds", rv.status == "ok" and values_path.exists(),
                f"status={rv.status!r}; message={rv.message!r}")
        c.check("formulas workbook builds", rf.status == "ok" and formulas_path.exists(),
                f"status={rf.status!r}; message={rf.message!r}")
        if not values_path.exists() or not formulas_path.exists():
            return

        typed = getattr(rv, "comparison_outcome", None)
        counts = getattr(typed, "counts", None)
        per_field = dict(getattr(counts, "per_field_counts", {}) or {})
        def field_count(label):
            matches = [value for key, value in per_field.items()
                       if key == label or key.endswith(f":{label}")]
            return matches[0] if len(matches) == 1 else None

        c.check("typed outcome keeps Boolean policy out of Text false positives",
                field_count("Text") == 9,
                f"per_field_counts={per_field!r}")
        c.check("typed outcome keeps exact large Med-Wid difference",
                field_count("Med-Wid") == 2,
                f"per_field_counts={per_field!r}")

        wv = load_workbook(values_path, data_only=False)
        wf = load_workbook(formulas_path, data_only=False)
        try:
            cv = wv["Comparison"]
            cf = wf["Comparison"]
            diffs_col_v = _header_col(cv, "Diffs")
            diffs_col_f = _header_col(cf, "Diffs")
            text_col_v = _header_col(cv, "Text")
            text_col_f = _header_col(cf, "Text")
            med_col_v = _header_col(cv, "Med-Wid")
            med_col_f = _header_col(cf, "Med-Wid")

            def value_at(key, col):
                return cv.cell(_row_for(cv, key), col).value

            c.check("values: actual Boolean folds and matches text TRUE",
                    value_at("BOOL", diffs_col_v) == 0 and
                    value_at("BOOL", text_col_v) == "TRUE",
                    f"Diffs={value_at('BOOL', diffs_col_v)!r}; "
                    f"Text={value_at('BOOL', text_col_v)!r}")
            c.check("values: integer 1 is not mistaken for Boolean TRUE",
                    value_at("INT_TRUE", diffs_col_v) == 1 and
                    value_at("INT_TRUE", text_col_v) == f"1{MARK}TRUE",
                    f"Diffs={value_at('INT_TRUE', diffs_col_v)!r}; "
                    f"Text={value_at('INT_TRUE', text_col_v)!r}")
            c.check("values: blank is visibly distinct from zero",
                    value_at("BLANK", diffs_col_v) == 1 and
                    value_at("BLANK", text_col_v) == f"(blank){MARK}0",
                    f"Diffs={value_at('BLANK', diffs_col_v)!r}; "
                    f"Text={value_at('BLANK', text_col_v)!r}")
            c.check("values: literal marker content does not increment Diffs",
                    value_at("MARK", diffs_col_v) == 0 and
                    value_at("MARK", text_col_v) == f"ALPHA{MARK}BETA",
                    f"Diffs={value_at('MARK', diffs_col_v)!r}; "
                    f"Text={value_at('MARK', text_col_v)!r}")
            c.check("values: adjacent large Med-Wid values stay different",
                    value_at("MED_PRECISE", diffs_col_v) == 1,
                    f"Diffs={value_at('MED_PRECISE', diffs_col_v)!r}")
            c.check("values: Med-Wid typed blank stays distinct from numeric zero",
                    value_at("MED_BLANK_ZERO", diffs_col_v) == 1 and
                    value_at("MED_BLANK_ZERO", med_col_v) == f"(blank){MARK}0",
                    f"Diffs={value_at('MED_BLANK_ZERO', diffs_col_v)!r}; "
                    f"Med-Wid={value_at('MED_BLANK_ZERO', med_col_v)!r}")
            c.check("values: >15-digit exact integer/text pair stays equal",
                    value_at("BIGEQ", diffs_col_v) == 0 and
                    value_at("BIGEQ", text_col_v) == "12345678901234567",
                    f"Diffs={value_at('BIGEQ', diffs_col_v)!r}; "
                    f"Text={value_at('BIGEQ', text_col_v)!r}")

            text_formula = str(cf.cell(_row_for(cf, "CASE"), text_col_f).value)
            med_formula = str(cf.cell(_row_for(cf, "MED_PRECISE"), med_col_f).value)
            blank_formula = str(cf.cell(_row_for(cf, "BLANK"), text_col_f).value)
            diffs_formula = str(cf.cell(_row_for(cf, "MARK"), diffs_col_f).value)

            state_cols_v = _state_columns(cv)
            state_cols_f = _state_columns(cf)
            state_failures = []
            if state_cols_v != state_cols_f or not state_cols_f:
                state_failures.append(("columns", state_cols_v, state_cols_f))
            for ws_name, ws, columns, formula_mode in (
                    ("values", cv, state_cols_v, False),
                    ("formulas", cf, state_cols_f, True)):
                for column in columns:
                    letter = ws.cell(1, column).column_letter
                    if not ws.column_dimensions[letter].hidden:
                        state_failures.append((ws_name, "visible", letter))
                for key, expected in expected_masks.items():
                    row = _row_for(ws, key)
                    cells = [ws.cell(row, column) for column in columns]
                    if formula_mode:
                        if any(cell.data_type != "f" for cell in cells):
                            state_failures.append((ws_name, key, "nonformula",
                                                   [(c.coordinate, c.data_type)
                                                    for c in cells]))
                        if any(len(str(cell.value)) > 8192 for cell in cells):
                            state_failures.append((ws_name, key, "too-long",
                                                   [len(str(c.value)) for c in cells]))
                    else:
                        actual = "".join(str(cell.value or "") for cell in cells)
                        if actual != expected:
                            state_failures.append((ws_name, key, expected, actual))
                        if any(cell.data_type == "f" for cell in cells):
                            state_failures.append((ws_name, key, "live values mask"))
                        want_diffs = None if "U" in expected else expected.count("D")
                        if ws.cell(row, diffs_col_v).value != want_diffs:
                            state_failures.append((ws_name, key, "Diffs",
                                                   want_diffs,
                                                   ws.cell(row, diffs_col_v).value))
            c.check("state twin: hidden versioned Comparison masks are live in the "
                    "formula workbook and exact literals in the values workbook",
                    not state_failures, repr(state_failures[:8]))

            state_col_f = state_cols_f[0] if state_cols_f else None
            state_formula = (str(cf.cell(_row_for(cf, "CASE"), state_col_f).value)
                             if state_col_f else "")
            one_sided_state_formula = (
                str(cf.cell(_row_for(cf, "LEFT_ONLY"), state_col_f).value)
                if state_col_f else "")

            helper_headers = [
                f"__CMP_E1_MW_V1_F002_{stage}"
                for stage in ("TRIM", "CORE", "VALID", "MASK", "CANON")
            ]
            helper_failures = []
            helper_formulas = []
            for book_name, book in (("values", wv), ("formulas", wf)):
                for side in ("LEFT", "RIGHT"):
                    ws = book[side]
                    expected_data_rows = (len(rows_left) if side == "LEFT"
                                          else len(rows_right))
                    try:
                        positions = [_header_col(ws, header)
                                     for header in helper_headers]
                    except Exception as exc:
                        helper_failures.append((book_name, side, type(exc).__name__,
                                                str(exc)))
                        continue
                    if positions != [6, 7, 8, 9, 10]:
                        helper_failures.append((book_name, side, "positions", positions))
                    if _header_col(ws, "Med-Wid") != 4 or _header_col(ws, "Key (helper)") != 5:
                        helper_failures.append((book_name, side, "visible geometry",
                                                _header_col(ws, "Med-Wid"),
                                                _header_col(ws, "Key (helper)")))
                    if ws.freeze_panes != "C2" or not str(ws.auto_filter.ref).startswith("A1:D"):
                        helper_failures.append((book_name, side, "view geometry",
                                                ws.freeze_panes, ws.auto_filter.ref))
                    for column in positions:
                        letter = ws.cell(1, column).column_letter
                        if not ws.column_dimensions[letter].hidden:
                            helper_failures.append((book_name, side, "visible helper", letter))
                        for row in range(2, expected_data_rows + 2):
                            cell = ws.cell(row, column)
                            if book_name == "formulas":
                                if cell.data_type != "f":
                                    helper_failures.append((book_name, side, "nonformula",
                                                            cell.coordinate,
                                                            cell.data_type))
                                    break
                                helper_formulas.append(str(cell.value))
                            elif cell.data_type == "f":
                                helper_failures.append((book_name, side,
                                                        "values helper is live",
                                                        cell.coordinate))
            c.check("formula twin: versioned Med-Wid helpers append after Key(helper) "
                    "with identical hidden geometry",
                    not helper_failures, repr(helper_failures[:5]))

            left_values = wv["LEFT"]
            space_row = _row_for(left_values, "SPACE")
            space_helpers = [left_values.cell(space_row, col).value
                             for col in range(6, 11)]
            c.check("values workbook stores exact literal Med-Wid stages (no hidden "
                    "recalculation load)",
                    space_helpers == ["0006.5000V", "0006.5000", True,
                                      "000XXX000", "6.5V"],
                    repr(space_helpers))

            forbidden_helpers = ("VALUE(", "NUMBERVALUE(", "DECIMAL(", "LET(",
                                 "LAMBDA(", "INDIRECT(")
            bad_helpers = [
                formula for formula in helper_formulas
                if len(formula) > 8192
                or any(token in formula.upper() for token in forbidden_helpers)
            ]
            c.check("formula twin: every Med-Wid helper is legacy, exact, and below "
                    "Excel's formula limit",
                    len(helper_formulas) == 2 * len(rows_left) * 5
                    and not bad_helpers,
                    f"helper_count={len(helper_formulas)}; expected="
                    f"{2 * len(rows_left) * 5}; bad={bad_helpers[:2]!r}")

            c.check("formula: ordinary text equality is case-sensitive via EXACT",
                    "EXACT(" in state_formula.upper()
                    and "INDEX(LEFT!C:C" in state_formula.upper()
                    and "INDEX(RIGHT!C:C" in state_formula.upper(),
                    state_formula)
            c.check("formula: ASCII-space policy still uses Excel TRIM",
                    "TRIM(" in state_formula.upper()
                    and "ISBLANK(INDEX(" in state_formula.upper(),
                    state_formula)
            forbidden_ws = ("CLEAN(", "CHAR(9)", "CHAR(10)", "CHAR(13)",
                            "CHAR(160)", "UNICHAR(160)")
            c.check("formula: no tab/CR/LF/NBSP global folding was introduced",
                    not any(token in state_formula.upper() for token in forbidden_ws),
                    state_formula)
            c.check("formula: Med-Wid suffix equality is case-sensitive via EXACT",
                    "EXACT(" in state_formula.upper()
                    and "INDEX(LEFT!J:J" in state_formula.upper()
                    and "INDEX(RIGHT!J:J" in state_formula.upper(),
                    state_formula)
            c.check("formula: one-sided rows carry explicit U codes in the mask",
                    'REPT("U",2)' in one_sided_state_formula,
                    one_sided_state_formula)
            spot_formulas_all = _formula_cells(wf["Spot Check"])
            spot_med_exact = [
                formula for formula in spot_formulas_all
                if "EXACT($X" in formula.upper() and "$AC" in formula.upper()
            ]
            spot_helper_failures = []
            spot_helper_formulas = []
            expected_spot_headers = [
                f"__CMP_E1_MW_V1_SPOT_{side}_{stage}"
                for side in ("A", "B")
                for stage in ("TRIM", "CORE", "VALID", "MASK", "CANON")
            ]
            for book_name, book in (("values", wv), ("formulas", wf)):
                ws = book["Spot Check"]
                actual_headers = [ws.cell(15, col).value for col in range(20, 30)]
                if actual_headers != expected_spot_headers:
                    spot_helper_failures.append((book_name, "headers", actual_headers))
                for col in range(20, 30):
                    letter = ws.cell(15, col).column_letter
                    if not ws.column_dimensions[letter].hidden:
                        spot_helper_failures.append((book_name, "visible", letter))
                for row in range(16, ws.max_row + 1):
                    for col in range(20, 30):
                        cell = ws.cell(row, col)
                        if cell.data_type == "f":
                            spot_helper_formulas.append(str(cell.value))
            c.check("formula: Spot Check independently stages both Med-Wid sides "
                    "and compares CANON with EXACT",
                    bool(spot_med_exact) and not spot_helper_failures
                    and len(spot_helper_formulas) == 20
                    and any("INDEX(LEFT!D:D" in formula.upper()
                            for formula in spot_helper_formulas)
                    and any("INDEX(RIGHT!D:D" in formula.upper()
                            for formula in spot_helper_formulas)
                    and all(len(formula) <= 8192
                            and not any(token in formula.upper()
                                        for token in forbidden_helpers)
                            for formula in spot_helper_formulas),
                    f"exact={spot_med_exact!r}; failures={spot_helper_failures[:3]!r}; "
                    f"helper_count={len(spot_helper_formulas)}")
            spot_state_failures = []
            expected_state_headers = [
                "__CMP_E1_STATE_V1_SPOT_INDEPENDENT_STATE",
                "__CMP_E1_STATE_V1_SPOT_EXPECTED_DISPLAY",
                "__CMP_E1_STATE_V1_SPOT_COMPARISON_STATE",
            ]
            for book_name, book in (("values", wv), ("formulas", wf)):
                ws = book["Spot Check"]
                actual = [ws.cell(15, col).value for col in range(11, 14)]
                if actual != expected_state_headers:
                    spot_state_failures.append((book_name, "headers", actual))
                for col in range(11, 14):
                    letter = ws.cell(15, col).column_letter
                    if not ws.column_dimensions[letter].hidden:
                        spot_state_failures.append((book_name, "visible", letter))
                    for row in (16, 17):
                        cell = ws.cell(row, col)
                        if cell.data_type != "f" or len(str(cell.value)) > 8192:
                            spot_state_failures.append(
                                (book_name, cell.coordinate, cell.data_type,
                                 len(str(cell.value))))
            spot_ws = wf["Spot Check"]
            spot_text_state = str(spot_ws["K16"].value)
            spot_med_state = str(spot_ws["K17"].value)
            spot_expected = str(spot_ws["L16"].value)
            spot_cmp_state = str(spot_ws["M16"].value)
            spot_cmp_display = str(spot_ws["F16"].value)
            spot_agree = str(spot_ws["G16"].value)
            c.check("formula: Spot Check independently recomputes state and the full "
                    "display, then EXACT-compares both twins",
                    not spot_state_failures
                    and "EXACT(" in spot_text_state.upper()
                    and "ISBLANK(INDEX(LEFT!C:C" in spot_text_state.upper()
                    and "EXACT($X17,$AC17)" in spot_med_state.upper()
                    and '$K16="D"' in spot_expected.upper()
                    and MARK in spot_expected
                    and "MID(INDEX(COMPARISON!$" in spot_cmp_state.upper()
                    and "ISBLANK(INDEX(COMPARISON!" in spot_cmp_display.upper()
                    and "EXACT($K16,$M16)" in spot_agree.upper()
                    and "EXACT($L16,$F16)" in spot_agree.upper(),
                    f"failures={spot_state_failures[:3]!r}; state={spot_text_state}; "
                    f"med={spot_med_state}; expected={spot_expected}; "
                    f"comparison_state={spot_cmp_state}; agree={spot_agree}")
            c.check("formula: Spot Med-Wid source wrapper preserves typed blank "
                    "before INDEX can coerce it to zero",
                    any("ISBLANK(INDEX(LEFT!D:D" in formula.upper()
                        for formula in spot_helper_formulas)
                    and any("ISBLANK(INDEX(RIGHT!D:D" in formula.upper()
                            for formula in spot_helper_formulas),
                    repr([formula for formula in spot_helper_formulas
                          if "INDEX(LEFT!D:D" in formula.upper()
                          or "INDEX(RIGHT!D:D" in formula.upper()][:2]))
            numeric_coercers = ("VALUE(", "NUMBERVALUE(", "DECIMAL(")
            c.check("formula: Med-Wid normalization uses no lossy numeric coercion",
                    not any(token in (state_formula + "\n" + med_formula + "\n" +
                                      "\n".join(helper_formulas +
                                                spot_helper_formulas +
                                                spot_med_exact)).upper()
                            for token in numeric_coercers),
                    state_formula + "\n" + "\n".join(spot_med_exact[:1]))
            c.check("formula: blank display is explicit and compares trimmed text",
                    "(blank)" in blank_formula and
                    "TRIM(" in blank_formula.upper(), blank_formula)
            c.check("formula: display still includes the familiar difference separator",
                    MARK in text_formula and "MID($" in text_formula.upper(),
                    text_formula)

            summary_formulas = _formula_cells(wf["Summary"])
            summary_scans = [f for f in summary_formulas
                             if _has_marker_state_scan(f)]
            spot_scans = [f for f in _formula_cells(wf["Spot Check"])
                          if _has_marker_state_scan(f)]
            cf_formulas = _cf_formulas(cf)
            cf_scans = [f for f in cf_formulas if _has_marker_state_scan(f)]
            c.check("formula: Comparison Diffs does not infer state from marker text",
                    not _has_marker_state_scan(diffs_formula)
                    and "SUBSTITUTE(" in diffs_formula.upper()
                    and '"D"' in diffs_formula
                    and MARK not in diffs_formula,
                    diffs_formula)
            c.check("formula: Summary does not infer state from marker text",
                    not summary_scans
                    and len([formula for formula in summary_formulas
                             if "MID(COMPARISON!" in formula.upper()
                             and '="D"' in formula.upper()]) == 2,
                    repr(summary_scans[:3] or summary_formulas))
            c.check("formula: Spot Check does not infer state from marker text",
                    not spot_scans, repr(spot_scans[:3]))
            c.check("formula: conditional formatting does not infer state from marker text",
                    not cf_scans
                    and any("MID($" in formula.upper() and '="D"' in formula.upper()
                            for formula in cf_formulas),
                    repr(cf_scans[:3] or cf_formulas))
            c.check("formula: hidden state formulas contain codes, not display separators",
                    all(MARK not in str(cf.cell(row, column).value)
                        for column in state_cols_f
                        for row in range(2, cf.max_row + 1)),
                    state_formula)

            bool_cell = _cell_for(wf, "LEFT", "BOOL", "Text")
            int_cell = _cell_for(wf, "LEFT", "INT_TRUE", "Text")
            big_cell = _cell_for(wf, "LEFT", "BIGEQ", "Text")
            boundary_cell = _cell_for(wf, "LEFT", "BOUNDARY", "Text")
            decimal_scale_cell = _cell_for(wf, "LEFT", "DEC_SCALE", "Text")
            decimal_exp_cell = _cell_for(wf, "LEFT", "DEC_EXP", "Text")
            float_small_cell = _cell_for(wf, "LEFT", "FLOAT_SMALL", "Text")
            float_large_cell = _cell_for(wf, "LEFT", "FLOAT_LARGE", "Text")
            c.check("writer: normalized Boolean is literal uppercase text",
                    bool_cell.data_type == "s" and bool_cell.value == "TRUE",
                    f"type={bool_cell.data_type!r}; value={bool_cell.value!r}")
            c.check("writer: integer 1 is exact numeric text, not Boolean text",
                    int_cell.data_type == "s" and int_cell.value == "1"
                    and int_cell.number_format == "@",
                    f"type={int_cell.data_type!r}; value={int_cell.value!r}")
            c.check("writer: >15-significant-digit integer is exact text",
                    big_cell.data_type == "s" and
                    big_cell.value == "12345678901234567",
                    f"type={big_cell.data_type!r}; value={big_cell.value!r}")
            c.check("writer: 15-digit boundary is exact text (no rendering seam)",
                    boundary_cell.data_type == "s" and
                    boundary_cell.value == "123456789012345"
                    and boundary_cell.number_format == "@",
                    f"type={boundary_cell.data_type!r}; value={boundary_cell.value!r}")
            numeric_text = {
                "decimal_scale": (decimal_scale_cell.value, "1.2300"),
                "decimal_exp": (decimal_exp_cell.value, "1E+3"),
                "float_small": (float_small_cell.value, "1e-07"),
                "float_large": (float_large_cell.value,
                                "100000000000000000000"),
            }
            c.check("writer: Decimal scale/exponent and float notation stay exact text",
                    all(actual == expected for actual, expected in numeric_text.values())
                    and all(cell.data_type == "s" and cell.number_format == "@"
                            for cell in (decimal_scale_cell, decimal_exp_cell,
                                         float_small_cell, float_large_cell)),
                    repr(numeric_text))

            error_types = []
            for i, token in enumerate(ERROR_CODES, start=1):
                cell = _cell_for(wf, "LEFT", f"ERR_{i}", "Text")
                if cell.data_type != "s" or cell.value != token:
                    error_types.append((token, cell.data_type, cell.value))
            c.check("writer: every literal Excel error remains exact text",
                    not error_types, repr(error_types))
        finally:
            wv.close()
            wf.close()
        if run_excel:
            _test_installed_excel(c, formulas_path, values_path, expected_masks)


def _test_installed_excel(c, formulas_path, values_path, expected_masks):
    """Optional real-Excel gate for cached formula results.

    The default check remains hermetic/openpyxl-only. ``--excel`` opts into COM,
    selects the Med-Wid blank-v-zero fixture in Spot Check, performs a full
    rebuild, saves cached values, and compares them to the literal twin.
    """
    print("\nInstalled Excel formula evaluation:")
    try:
        import win32com.client as win32
    except Exception as exc:
        c.check("installed Excel COM dependency is available", False,
                f"{type(exc).__name__}: {exc}")
        return

    inspection = load_workbook(formulas_path, data_only=False, read_only=False)
    try:
        comparison = inspection["Comparison"]
        med_blank_row = _row_for(comparison, "MED_BLANK_ZERO")
    finally:
        inspection.close()

    excel = book = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        book = excel.Workbooks.Open(str(Path(formulas_path).resolve()),
                                    UpdateLinks=0, ReadOnly=False)
        book.Worksheets("Spot Check").Range("C6").Value = med_blank_row
        excel.CalculateFullRebuild()
        book.Save()
        book.Close(SaveChanges=False)
        book = None
        c.check("installed Excel full rebuild and save completes", True)
    except Exception as exc:
        c.check("installed Excel full rebuild and save completes", False,
                f"{type(exc).__name__}: {exc}")
        return
    finally:
        if book is not None:
            try:
                book.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass

    calculated = load_workbook(formulas_path, data_only=True)
    literal = load_workbook(values_path, data_only=True)
    try:
        cf = calculated["Comparison"]
        cv = literal["Comparison"]
        state_cols_f = _state_columns(cf)
        state_cols_v = _state_columns(cv)
        diffs_f, diffs_v = _header_col(cf, "Diffs"), _header_col(cv, "Diffs")
        text_f, text_v = _header_col(cf, "Text"), _header_col(cv, "Text")
        med_f, med_v = _header_col(cf, "Med-Wid"), _header_col(cv, "Med-Wid")
        parity_failures = []
        for key, expected in expected_masks.items():
            row_f, row_v = _row_for(cf, key), _row_for(cv, key)
            actual_mask = "".join(
                str(cf.cell(row_f, column).value or "")
                for column in state_cols_f)
            literal_mask = "".join(
                str(cv.cell(row_v, column).value or "")
                for column in state_cols_v)
            if actual_mask != expected or literal_mask != expected:
                parity_failures.append((key, "mask", expected,
                                        actual_mask, literal_mask))
            want_diffs = None if "U" in expected else expected.count("D")
            actual_diffs = cf.cell(row_f, diffs_f).value
            literal_diffs = cv.cell(row_v, diffs_v).value
            if want_diffs is None:
                if actual_diffs not in (None, "") or literal_diffs not in (None, ""):
                    parity_failures.append((key, "Diffs", want_diffs,
                                            actual_diffs, literal_diffs))
            elif actual_diffs != want_diffs or literal_diffs != want_diffs:
                parity_failures.append((key, "Diffs", want_diffs,
                                        actual_diffs, literal_diffs))
            for field_name, formula_col, value_col in (
                    ("Text", text_f, text_v), ("Med-Wid", med_f, med_v)):
                actual = cf.cell(row_f, formula_col).value
                expected_display = cv.cell(row_v, value_col).value
                if actual != expected_display:
                    parity_failures.append((key, field_name,
                                            expected_display, actual))
        c.check("installed Excel: formula masks, Diffs, and displays equal the "
                "literal twin for every fixture row",
                not parity_failures, repr(parity_failures[:8]))

        summary = calculated["Summary"]

        def summary_value(label, column):
            for row in range(1, summary.max_row + 1):
                if summary.cell(row, 2).value == label:
                    value = summary.cell(row, column).value
                    if value is not None:
                        return value
            raise KeyError((label, column))

        summary_actual = {
            "total": summary_value("Total differing cells", 3),
            "Text": summary_value("Text", 4),
            "Med-Wid": summary_value("Med-Wid", 4),
        }
        c.check("installed Excel: Summary headline and per-field counts consume "
                "state masks correctly",
                summary_actual == {"total": 11, "Text": 9, "Med-Wid": 2},
                repr(summary_actual))

        spot = calculated["Spot Check"]
        blankish = lambda value: value in (None, "")
        spot_actual = {
            "text_state": spot["K16"].value,
            "text_cmp_state": spot["M16"].value,
            "text_agree": spot["G16"].value,
            "med_state": spot["K17"].value,
            "med_cmp_state": spot["M17"].value,
            "med_expected": spot["L17"].value,
            "med_display": spot["F17"].value,
            "med_agree": spot["G17"].value,
            "med_left_canon": spot["H17"].value,
            "med_right_canon": spot["I17"].value,
        }
        c.check("installed Excel: Spot Check catches Med-Wid blank-v-zero with "
                "independent state and exact full-display agreement",
                spot_actual["text_state"] == "E"
                and spot_actual["text_cmp_state"] == "E"
                and spot_actual["text_agree"] == "OK"
                and spot_actual["med_state"] == "D"
                and spot_actual["med_cmp_state"] == "D"
                and spot_actual["med_expected"] == f"(blank){MARK}0"
                and spot_actual["med_display"] == f"(blank){MARK}0"
                and spot_actual["med_agree"] == "OK"
                and blankish(spot_actual["med_left_canon"])
                and str(spot_actual["med_right_canon"]) == "0",
                repr(spot_actual))
    finally:
        calculated.close()
        literal.close()


def test_state_chunk_planning(c):
    """Wide schemas chunk deterministically and reject an impossible field."""
    fields = [f"F{i}" for i in range(1, 121)]
    schema = cc.CompareSchema(
        report_name="State-mask chunking",
        header=["Loc", *fields],
        side_a="LEFT",
        side_b="RIGHT",
        id_noun="row",
        id_noun_plural="rows",
    )
    layout = cc._Layout(schema, False)
    flattened = tuple(field for chunk in layout.state_chunks
                      for field in chunk["fields"])
    formulas = [cc._state_chunk_formula(layout, cc.XL_MAX_ROWS, chunk["fields"])
                for chunk in layout.state_chunks]
    c.check("state masks: wide schemas split into ordered, versioned chunks below "
            "Excel's formula ceiling",
            len(layout.state_chunks) > 1
            and flattened == tuple(layout.field_indices)
            and len({chunk["header"] for chunk in layout.state_chunks})
                    == len(layout.state_chunks)
            and all(chunk["header"].startswith("__CMP_E1_STATE_V1_C")
                    for chunk in layout.state_chunks)
            and all(len(formula) <= 8192 for formula in formulas)
            and layout.comparison_physical_n_cols
                    == layout.state_chunks[-1]["col_idx"],
            f"chunks={len(layout.state_chunks)}; lengths="
            f"{[len(formula) for formula in formulas]!r}")

    impossible = None
    with patch(cc, "_field_state_expr",
               lambda _lay, _row, _field: "X" * 8_193):
        try:
            cc._Layout(schema, False)
        except ValueError as exc:
            impossible = str(exc)
    c.check("state masks: a single formula beyond 8,192 characters hard-fails",
            impossible is not None and "Excel allows at most 8,192" in impossible,
            repr(impossible))


def test_state_mask_physical_column_limit(c):
    """State columns count even when every visible/data column still fits."""
    n_fields = 16_378  # 6 identity + 16,378 visible fields == Excel's 16,384 cap
    fields = [f"F{i}" for i in range(1, n_fields + 1)]
    schema = cc.CompareSchema(
        report_name="State-mask physical width",
        header=["Loc", *fields],
        side_a="LEFT",
        side_b="RIGHT",
        id_noun="row",
        id_noun_plural="rows",
    )
    row = ["K", *("same" for _ in fields)]
    with temp_dir("tsmis_state_width_") as tmp:
        path = Path(tmp) / "must-not-write.xlsx"
        result = cc.run_compare(schema, [row], [list(row)], False, path,
                                mode="formulas")
        c.check("physical state-mask width fails cleanly when visible/data widths "
                "would otherwise fit exactly",
                result.status == "error"
                and "past Excel's 16,384-column limit" in result.message
                and not path.exists(),
                f"status={result.status!r}; message={result.message!r}; "
                f"exists={path.exists()}")


def test_medwid_physical_column_limit(c):
    """Hidden stages count toward Excel's physical 16,384-column ceiling."""
    n_fields = 2_731                 # flat data sheet: 6*n + 3 = 16,389 columns
    fields = [f"MW{i}" for i in range(1, n_fields + 1)]
    schema = cc.CompareSchema(
        report_name="Med-Wid physical width",
        header=["Loc", *fields],
        side_a="LEFT",
        side_b="RIGHT",
        id_noun="row",
        id_noun_plural="rows",
        medwid_fields=tuple(fields),
    )
    row = ["K", *("06V" for _ in fields)]
    with temp_dir("tsmis_medwid_width_") as tmp:
        path = Path(tmp) / "must-not-write.xlsx"
        result = cc.run_compare(schema, [row], [list(row)], False, path,
                                mode="formulas")
        c.check("physical Med-Wid helper width fails cleanly before workbook write",
                result.status == "error"
                and "16,389 columns" in result.message
                and not path.exists(),
                f"status={result.status!r}; message={result.message!r}; "
                f"exists={path.exists()}")


def test_extreme_schema_fails_before_openpyxl_column_conversion(c):
    """A schema beyond openpyxl's address range still returns a typed error."""
    n_fields = 4_000
    fields = [f"MW{i}" for i in range(1, n_fields + 1)]
    schema = cc.CompareSchema(
        report_name="Extreme physical width",
        header=["Loc", *fields],
        side_a="LEFT",
        side_b="RIGHT",
        id_noun="row",
        id_noun_plural="rows",
        medwid_fields=tuple(fields),
    )
    row = ["K", *("06V" for _ in fields)]
    with temp_dir("tsmis_extreme_width_") as tmp:
        path = Path(tmp) / "must-not-write.xlsx"
        try:
            result = cc.run_compare(schema, [row], [list(row)], False, path,
                                    mode="formulas")
            raised = None
        except Exception as exc:
            result, raised = None, exc
        c.check("extreme helper width returns a clean Excel-limit result",
                raised is None and result.status == "error"
                and "past Excel's 16,384-column limit" in result.message
                and not path.exists(),
                f"raised={raised!r}; result={result!r}")

    # Visible/raw columns fit, but long escaped sheet labels + ditto formulas
    # force enough state chunks that the planned hidden index exceeds both XFD
    # and openpyxl's internal 18,278-column address ceiling. Planning must reject
    # numerically before get_column_letter sees that index.
    n_fields = 16_000
    fields = [f"F{i}" for i in range(1, n_fields + 1)]
    schema = cc.CompareSchema(
        report_name="Extreme state width",
        header=["Loc", *fields],
        side_a="A''''''''''''''''''''A",
        side_b="B''''''''''''''''''''B",
        id_noun="row",
        id_noun_plural="rows",
        ditto_nonasserting=True,
    )
    row = ["K", *("same" for _ in fields)]
    with temp_dir("tsmis_extreme_state_width_") as tmp:
        path = Path(tmp) / "must-not-write.xlsx"
        try:
            result = cc.run_compare(schema, [row], [list(row)], False, path,
                                    mode="formulas")
            raised = None
        except Exception as exc:
            result, raised = None, exc
        c.check("extreme state-chunk width rejects before openpyxl index conversion",
                raised is None and result.status == "error"
                and "past Excel's 16,384-column limit" in result.message
                and not path.exists(),
                f"raised={raised!r}; result={result!r}")


def test_nonfinite_numeric_rejected(c):
    """NaN/infinity must never become deceptive blank workbook cells."""
    print("\nNon-finite numeric fail-closed boundary:")
    cases = (
        ("float NaN", float("nan")),
        ("float +infinity", float("inf")),
        ("float -infinity", float("-inf")),
        ("Decimal NaN", Decimal("NaN")),
        ("Decimal +infinity", Decimal("Infinity")),
        ("Decimal -infinity", Decimal("-Infinity")),
    )
    with temp_dir("tsmis_nonfinite_policy_") as tmp:
        for index, (label, value) in enumerate(cases):
            path = Path(tmp) / f"nonfinite-{index}.xlsx"
            values_path = path.with_name(f"{path.stem} (values){path.suffix}")
            result = cc.run_compare(
                TEXT_SCHEMA,
                [["K", value]],
                [["K", str(value)]],
                False,
                path,
                mode="both",
            )
            typed = getattr(result, "comparison_outcome", None)
            c.check(
                f"{label} is rejected before either workbook is written",
                result.status == "error"
                and "non-finite numeric value" in result.message
                and not path.exists() and not values_path.exists()
                and getattr(typed, "completion", None) == "failed",
                f"status={result.status!r}; message={result.message!r}; "
                f"typed={typed!r}",
            )


def test_literal_helper_numeric_scope(c):
    """Source exactness must not textify engine-owned numeric counters."""
    print("\nLiteral-cell numeric scope:")
    wb = Workbook()
    try:
        ws = wb.active
        small = cc.set_safe_literal_cell(ws["A1"], 7)
        large = cc.set_safe_literal_cell(ws["A2"], 12345678901234567)
        source = cc.set_safe_literal_cell(
            ws["A3"], Decimal("1.2300"), exact_source_numeric=True)
        c.check("default guard keeps small engine counters numeric",
                small.data_type == "n" and small.value == 7,
                f"type={small.data_type!r}; value={small.value!r}")
        c.check("default guard retains the >15-significant-digit backstop",
                large.data_type == "s" and large.value == "12345678901234567",
                f"type={large.data_type!r}; value={large.value!r}")
        c.check("source-only numeric mode preserves Decimal lexical scale",
                source.data_type == "s" and source.value == "1.2300"
                and source.number_format == "@",
                f"type={source.data_type!r}; value={source.value!r}")
    finally:
        wb.close()


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--excel", action="store_true",
        help=("also open the formulas workbook in installed Microsoft Excel, "
              "perform CalculateFullRebuild, and verify cached results against "
              "the values twin"))
    args = parser.parse_args(argv)
    c = Checker()
    test_python_policy(c)
    test_workbook_policy(c, run_excel=args.excel)
    test_state_chunk_planning(c)
    test_state_mask_physical_column_limit(c)
    test_medwid_physical_column_limit(c)
    test_extreme_schema_fails_before_openpyxl_column_conversion(c)
    test_nonfinite_numeric_rejected(c)
    test_literal_helper_numeric_scope(c)
    raise SystemExit(c.summary())


if __name__ == "__main__":
    main()
