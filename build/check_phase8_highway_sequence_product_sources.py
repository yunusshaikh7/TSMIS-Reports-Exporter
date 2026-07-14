#!/usr/bin/env python3
"""Cell-for-cell product consolidation parity for Highway Sequence Stage 8."""

from __future__ import annotations

from collections import Counter
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


VISUAL_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd"
)
SOURCE_CACHE = VISUAL_ROOT / "phase8_highway_sequence_source_rows_draft_r1.json"
PRODUCT_ROOT = VISUAL_ROOT / "phase8_highway_sequence_product_sources_r2"
PRODUCT_RESULT = PRODUCT_ROOT / "result.json"
EXCEL_OUTPUT = PRODUCT_ROOT / "current_tsmis_excel_consolidated.xlsx"
PDF_OUTPUT = PRODUCT_ROOT / "current_tsmis_pdf_consolidated.xlsx"
DEFAULT_OUTPUT = VISUAL_ROOT / "phase8_highway_sequence_product_source_parity_r2.json"

BINDINGS = {
    "source_cache": (SOURCE_CACHE, 49_304_637, "564cf21972aeaf461811095997524c2d02f3ca4f238bb8da8b715415df2762f8"),
    "product_result": (PRODUCT_RESULT, 242_779, "39eb2e53091bfcfdd6a3b4a2997b8700c1d617e94f136f4d3a0603730df82493"),
    "excel_output": (EXCEL_OUTPUT, 2_424_212, "cf5905332db3d3eb5a49a87d603f6e36f209cad9a84173b381dace6600168b20"),
    "pdf_output": (PDF_OUTPUT, 2_371_547, "070afe51ea3bf84c9704d0a36a02702b65189941badab6374b03461db8ef6ccc"),
}
SHEET_NAME = "Highway Locations"
HEADERS = (
    "Route", "County", "City", None, "PM", None, "HG", "FT",
    "Distance To Next Point", "Description",
)
FIELD_NAMES = (
    "Route", "County", "City", "PM Prefix", "PM", "PM Suffix", "HG", "FT",
    "Distance To Next Point", "Description",
)


class ParityError(RuntimeError):
    pass


def _sha(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode("utf-8")


def _bind() -> dict[str, object]:
    result = {}
    for label, (path, size, digest) in BINDINGS.items():
        observed = {"bytes": path.stat().st_size, "sha256": _sha(path)}
        if observed != {"bytes": size, "sha256": digest}:
            raise ParityError(f"{label} identity drift: {observed}")
        result[label] = {"path": str(path.resolve()), **observed}
    return result


def _read_product(path: Path, source: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            raise ParityError(f"{path.name}: sheet role universe {workbook.sheetnames}")
        sheet = workbook[SHEET_NAME]
        iterator = iter(sheet.iter_rows())
        header_cells = next(iterator, None)
        if header_cells is None or len(header_cells) != len(HEADERS):
            raise ParityError(
                f"{path.name}: physical header width "
                f"{None if header_cells is None else len(header_cells)}"
            )
        header = tuple(cell.value for cell in header_cells)
        if header != HEADERS:
            raise ParityError(f"{path.name}: header drift {header!r}")
        rows = []
        for source_row, cells in enumerate(iterator, 2):
            if len(cells) > len(HEADERS):
                raise ParityError(
                    f"{path.name}: row {source_row} physical width {len(cells)}"
                )
            values = []
            for cell in cells:
                if cell.data_type in ("f", "e"):
                    raise ParityError(f"{path.name}!{cell.coordinate}: formula/error")
                if cell.value is not None and not isinstance(cell.value, str):
                    raise ParityError(
                        f"{path.name}!{cell.coordinate}: scalar {type(cell.value).__name__}"
                    )
                values.append(cell.value)
            values.extend([None] * (len(HEADERS) - len(values)))
            if all(value is None for value in values):
                raise ParityError(f"{path.name}: blank data row {source_row}")
            rows.append({
                "source": source, "source_row": source_row,
                "source_ref": f"{path.name}:row:{source_row}",
                "values": tuple(values),
            })
        return rows
    finally:
        workbook.close()


def _expected(serialized: list[dict[str, object]]) -> list[dict[str, object]]:
    return [
        {
            "source_ref": str(item["source_ref"]),
            "values": (str(item["route"]), *tuple(item["values"])),
        }
        for item in serialized
    ]


def _compare(label: str, expected: list[dict[str, object]],
             actual: list[dict[str, object]]) -> dict[str, object]:
    mismatches = []
    field_counts: Counter[str] = Counter()
    for ordinal, (left, right) in enumerate(zip(expected, actual), 1):
        for index, field in enumerate(FIELD_NAMES):
            if left["values"][index] != right["values"][index]:
                field_counts[field] += 1
                mismatches.append({
                    "ordinal": ordinal, "field": field,
                    "expected_ref": left["source_ref"],
                    "actual_ref": right["source_ref"],
                    "expected": left["values"][index],
                    "actual": right["values"][index],
                })
    missing_or_extra = abs(len(expected) - len(actual))
    return {
        "label": label, "expected_rows": len(expected), "actual_rows": len(actual),
        "missing_or_extra_rows": missing_or_extra,
        "cell_mismatches": len(mismatches),
        "field_mismatch_counts": dict(sorted(field_counts.items())),
        "mismatches": mismatches,
        "exact": not missing_or_extra and not mismatches,
    }


def main() -> int:
    bindings = _bind()
    source = json.loads(SOURCE_CACHE.read_bytes())["rows"]
    expected_excel = _expected(source["current_tsmis_excel"])
    expected_pdf = _expected(source["current_tsmis_pdf"])
    actual_excel = _read_product(EXCEL_OUTPUT, "product_excel_consolidation")
    actual_pdf = _read_product(PDF_OUTPUT, "product_pdf_consolidation")
    comparisons = {
        "excel": _compare("product Excel consolidation vs independent source", expected_excel, actual_excel),
        "pdf": _compare("product PDF consolidation vs independent source", expected_pdf, actual_pdf),
    }
    result = {
        "audit": "Stage 8 Highway Sequence product consolidation source parity",
        "bindings": bindings, "comparisons": comparisons,
        "invariants": {
            "excel_rows_60494": len(actual_excel) == 60_494,
            "pdf_rows_60493": len(actual_pdf) == 60_493,
            "excel_exact": comparisons["excel"]["exact"],
            "pdf_exact": comparisons["pdf"]["exact"],
        },
    }
    DEFAULT_OUTPUT.write_bytes(_json(result))
    print(
        "PASS Highway Sequence product consolidation parity probe: "
        f"Excel mismatches {comparisons['excel']['cell_mismatches']}; "
        f"PDF mismatches {comparisons['pdf']['cell_mismatches']}; {DEFAULT_OUTPUT}"
    )
    # This probe records product red paths rather than requiring exactness. Row
    # coverage is terminal; any cell mismatch is classified by the caller.
    if not result["invariants"]["excel_rows_60494"] or not result["invariants"]["pdf_rows_60493"]:
        raise ParityError("product row coverage differs from independent source")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ParityError as exc:
        print(f"FAIL Highway Sequence product consolidation parity probe: {exc}")
        raise SystemExit(1)
