"""Golden check for SHEET-FORMULA-INJECTION (compare_core + the consolidators).

A free-text value that begins with = + - @ (e.g. a malicious Description like
"=cmd|'/C calc'!A1") was written straight through; openpyxl turns a leading "="
into a live FORMULA, so opening the workbook in Excel could execute it. The
shared guard (compare_core.is_formula_injection + the per-writer data_type='s'
forcing) keeps the value byte-for-byte but stores it as TEXT, so Excel shows it
verbatim and never runs it. The same guard covers openpyxl's seven ERROR_CODES,
which otherwise become live error cells. Clean data is unaffected (the
regression lock holds — proven separately by the cell-dump diff).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_injection.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from compare_core import CompareSchema, is_formula_injection, run_compare
from consolidate_xlsx_base import consolidate_xlsx
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ERROR_CODES

EVIL = '=HYPERLINK("http://x",1)'
EVIL2 = "=cmd|'/C calc'!A1"
SC = CompareSchema(report_name="Inj", header=["Loc", "Desc"], side_a="AENV",
                   side_b="BENV", id_noun="row", id_noun_plural="rows",
                   sides_noun="environments")
ERROR_ROWS = [[f"E{i}", value] for i, value in enumerate(ERROR_CODES, start=1)]
ROWS_A = [["1.0", EVIL], ["2.0", "normal"], ["3.0", EVIL2]] + ERROR_ROWS
ROWS_B = [["1.0", EVIL], ["2.0", "normal"]] + ERROR_ROWS


def _col_of(ws, header_name):
    for c in next(ws.iter_rows(max_row=1)):
        if c.value == header_name:
            return c.column            # 1-based index
    raise AssertionError(f"{header_name!r} not in {ws.title}")


def _cell_type(ws, header_name, want_value):
    col = _col_of(ws, header_name)
    for row in ws.iter_rows(min_row=2):
        c = row[col - 1]
        if c.value == want_value:
            return c.data_type
    raise AssertionError(f"{want_value!r} not found under {header_name!r} in {ws.title}")


def test_predicate():
    for v in (EVIL, EVIL2, "+1", "-x", "@y", *ERROR_CODES):
        assert is_formula_injection(v), v
    for v in ("normal", "1.0", "#N/A with context", 5, None, "(blank)"):
        assert not is_formula_injection(v), v


def test_compare_values_flavor():
    out = os.path.join(tempfile.gettempdir(), "_inj_cmp.xlsx")
    run_compare(SC, ROWS_A, ROWS_B, False, out, mode="values")
    wb = load_workbook(out, data_only=False)
    # data sheet: raw "=HYPERLINK…" stored as text, not a formula
    assert _cell_type(wb["AENV"], "Desc", EVIL) == "s", "data-sheet Desc not guarded"
    # Comparison: matched value shown verbatim, guarded
    assert _cell_type(wb["Comparison"], "Desc", EVIL) == "s", "Comparison Desc not guarded"
    # Only-in: the A-only evil row, guarded
    assert _cell_type(wb["Only in AENV"], "Desc", EVIL2) == "s", "Only-in Desc not guarded"
    for value in ERROR_CODES:
        assert _cell_type(wb["AENV"], "Desc", value) == "s", value
        assert _cell_type(wb["BENV"], "Desc", value) == "s", value
        assert _cell_type(wb["Comparison"], "Desc", value) == "s", value
    wb.close()
    os.remove(out)


def test_compare_formulas_flavor():
    out = os.path.join(tempfile.gettempdir(), "_inj_cmp_f.xlsx")
    run_compare(SC, ROWS_A, ROWS_B, False, out, mode="formulas")
    wb = load_workbook(out, data_only=False)
    # data sheet still guarded; the field cell is our own =IF formula (type f).
    assert _cell_type(wb["AENV"], "Desc", EVIL) == "s", "data-sheet Desc not guarded"
    for value in ERROR_CODES:
        assert _cell_type(wb["AENV"], "Desc", value) == "s", value
        assert _cell_type(wb["BENV"], "Desc", value) == "s", value
    g = _col_of(wb["Comparison"], "Desc")
    field = wb["Comparison"].cell(row=2, column=g)
    assert field.data_type == "f" and str(field.value).startswith("=IF"), \
        ("Comparison field must stay our formula", field.data_type, field.value)
    wb.close()
    os.remove(out)


def test_consolidate_streaming_and_normal():
    # consolidate_xlsx (streaming write path). The INPUT must carry the evil
    # value as TEXT — that's how a real TSMIS export stores a Description — so
    # force the cell to a string (a plain append would make it a live formula
    # with no cached value, which data_only reads back as blank).
    d = Path(tempfile.mkdtemp())
    wb = Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Loc", "Desc"]); ws.append(["1.0", EVIL2]); ws.append(["2.0", "ok"])
    ws.cell(row=2, column=2).data_type = "s"      # store EVIL2 as text in the input
    wb.save(d / "x_route_001.xlsx")
    out = d / "consolidated.xlsx"
    res = consolidate_xlsx(input_dir=d, out_path=out, sheet_name="S",
                           report_name="R", title="t")
    assert res.status == "ok", res.status
    wb = load_workbook(out)
    assert _cell_type(wb["S"], "Desc", EVIL2) == "s", "consolidate Desc not guarded"
    wb.close()

    # normal-mode override technique (as used by the TSN / Ramp Summary writers)
    p = d / "normal.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append([EVIL2])
    for c in ws[ws.max_row]:
        if is_formula_injection(c.value):
            c.data_type = "s"
    wb.save(p)
    wb = load_workbook(p)
    assert wb.active.cell(row=1, column=1).data_type == "s", \
        "normal-mode data_type override failed"
    wb.close()


def main():
    test_predicate()
    test_compare_values_flavor()
    test_compare_formulas_flavor()
    test_consolidate_streaming_and_normal()
    print("OK  SHEET-LITERAL-GUARD: leading =+-@ free text and all openpyxl "
          "ERROR_CODES are stored as TEXT on source-data write paths (data "
          "sheets, Comparison + Only-in, consolidators, normal-mode); our own "
          "formulas stay formulas.")


if __name__ == "__main__":
    main()
