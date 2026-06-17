"""Golden check for the cross-environment RAMP SUMMARY comparison
(scripts/compare_env.RAMP_SUMMARY + compare_core).

The 2026-06-16 audit validated the route-keyed cross-env Ramp Summary on real
3-environment data (PROD-vs-TEST = 32 genuine diff cells / 9 routes; PROD==ARS).
This pins that wiring with a planted-difference fixture so a future change can't
silently break it:

  * one route present on BOTH sides with ONE field changed  -> 1 differing cell,
  * one route present only in side A and one only in side B -> two one-sided rows,
  * a route whose number is written unpadded on one side ("5") and zero-padded on
    the other ("005") must still pair into ONE matched route, not split into two
    one-sided rows (compare_env._norm_route_key — a real PDF-title vs filename
    mismatch the audit called out).

Ramp Summary parses per-route PDFs, so parse_pdf is monkeypatched to hand the
adapter parsed records directly (same approach as check_ramp_summary_partial);
everything else — the loader, route normalization, schema and engine — runs for
real through compare_folders, and the written VALUES workbook is read back.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_ramp_summary.py
"""
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
import consolidate_ramp_summary as rs
from compare_core import _DIFF_MARK
from events import Events
from openpyxl import load_workbook

# (side, filename-route-token) -> the record parse_pdf would return. total_ramps
# is set everywhere so record_has_data() is True (a real parse); only route 008
# differs across sides (10 vs 11) — exactly one planted differing cell. Route 5
# is written unpadded on PROD and zero-padded on TEST to exercise the route-key
# normalizer; both must normalize to "005" and pair.
RECORDS = {
    ("prod", "5"):   {"route": "5",   "total_ramps": 20},
    ("prod", "8"):   {"route": "8",   "total_ramps": 10},
    ("prod", "10"):  {"route": "10",  "total_ramps": 15},
    ("prod", "99"):  {"route": "99",  "total_ramps": 5},
    ("test", "005"): {"route": "005", "total_ramps": 20},   # pairs with PROD "5"
    ("test", "8"):   {"route": "8",   "total_ramps": 11},   # planted 1-cell diff
    ("test", "10"):  {"route": "10",  "total_ramps": 15},
    ("test", "110"): {"route": "110", "total_ramps": 7},
}
PROD_TOKENS = ["5", "8", "10", "99"]
TEST_TOKENS = ["005", "8", "10", "110"]


def _fake_parse(path):
    s = str(path).replace("\\", "/").lower()
    side = "prod" if "ssor-prod" in s else "test"
    token = re.search(r"_route_(\w+)\.pdf$", Path(path).name).group(1)
    rec = dict(RECORDS[(side, token)])
    rec["source_file"] = Path(path).name
    return rec


def test_config_is_route_keyed():
    rsmod = compare_env.RAMP_SUMMARY
    assert rsmod.sheet_name is None, ("Ramp Summary is the PDF path", rsmod.sheet_name)
    assert rsmod.subdir == "ramp_summary", rsmod.subdir
    assert rsmod.key_col is None, ("route is the first column / the key", rsmod.key_col)
    assert rsmod.base_schema.id_noun == "route", rsmod.base_schema.id_noun


def test_planted_diff_and_one_sided_routes():
    root = Path(tempfile.mkdtemp())
    saved = (rs.parse_pdf, getattr(rs, "_DEPS_OK", False))
    rs.parse_pdf = _fake_parse
    rs._DEPS_OK = True                     # parse_pdf is stubbed; no pdfplumber needed
    try:
        a = root / "2026-06-16 ssor-prod" / "ramp_summary"
        b = root / "2026-06-16 ssor-test" / "ramp_summary"
        a.mkdir(parents=True)
        b.mkdir(parents=True)
        for t in PROD_TOKENS:
            (a / f"ramp_summary_route_{t}.pdf").write_bytes(b"%PDF-1.4")
        for t in TEST_TOKENS:
            (b / f"ramp_summary_route_{t}.pdf").write_bytes(b"%PDF-1.4")

        out = root / "cmp.xlsx"
        res = compare_env.RAMP_SUMMARY.compare_folders(
            a.parent, b.parent, out, events=Events(),
            confirm_overwrite=lambda _p: True, mode="values")
        assert res.status == "ok", (res.status, res.message)
        assert res.verdict == "diff", res.verdict
        assert "DIFFERENCES FOUND" in res.summary_lines[0], res.summary_lines[0]

        wb = load_workbook(out, read_only=True, data_only=True)
        body = list(wb["Comparison"].iter_rows(values_only=True))[1:]
        wb.close()
        # per-route shape (has_route=False): A=Route B=# C=A Row D=B Row
        #                                    E=Status F=Diffs G..fields
        routes = [r[0] for r in body]
        statuses = [r[4] for r in body]
        # Union routes: 005, 008, 010 (both) + 099 (PROD only) + 110 (TEST only).
        assert len(body) == 5, ("five union routes", routes)
        assert statuses.count("Both") == 3, statuses
        one_sided = [s for s in statuses if s and s != "Both"]
        assert len(one_sided) == 2, ("one PROD-only + one TEST-only", one_sided)

        # Route-key normalization: unpadded "5" and padded "005" became ONE
        # matched route, not two one-sided rows.
        assert routes.count("005") == 1, ("5 / 005 must pair into one route", routes)
        i005 = routes.index("005")
        assert statuses[i005] == "Both", ("the normalized route matches", statuses[i005])

        # Exactly one planted differing cell (Total Ramps on route 008).
        neq = sum(1 for r in body for v in r
                  if isinstance(v, str) and _DIFF_MARK in v)
        assert neq == 1, ("one planted Total Ramps difference on route 008", neq)
    finally:
        rs.parse_pdf, rs._DEPS_OK = saved
        shutil.rmtree(root, ignore_errors=True)


def main():
    test_config_is_route_keyed()
    test_planted_diff_and_one_sided_routes()
    print("OK  COMPARE-RAMP-SUMMARY: route-keyed cross-env compare pairs "
          "padded/unpadded route numbers, reports exactly the planted 1-cell "
          "difference, and pulls the two one-sided routes, end to end through "
          "compare_folders.")


if __name__ == "__main__":
    main()
