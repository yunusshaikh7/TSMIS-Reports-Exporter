#!/usr/bin/env python3
"""Independent Stage-8 Ramp Summary raw-to-comparison oracle.

Authority flows from exact All Reports 7.9 TSMIS PDF/Excel members and the exact
accepted TSN Ramp Summary chain.  This module imports no application parser,
normalizer, schema, comparator, or writer.  Production execution happens in a
separate child process and its workbooks are parsed back against independently
derived truth.

The outcomes are intentionally distinct:

* ``stage8_base_oracle_complete``: every source, route, category, comparison
  row, and product-output discrepancy has an exact disposition.
* ``production_value_projection_exact``: all values the product does emit are
  faithful to the raw sources.
* ``production_comparison_semantics_exact``: side taxonomy and verdict inputs
  match the approved comparison contract.  This remains false until the known
  Ramp Summary product findings are remediated in Stage 11.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
import hashlib
import json
import logging
import os
from pathlib import Path
import re
import subprocess
import sys
import tempfile
from typing import Iterable, Sequence
import zipfile

import pdfplumber
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter
import pypdf
from pypdf import PdfReader


REPO_ROOT = Path(__file__).resolve().parent.parent
GENERATOR_PATH = Path(__file__).resolve()
SELF_GATE_PATH = GENERATOR_PATH.with_name("check_phase8_ramp_summary_comparison.py")
PRODUCT_HELPER_PATH = GENERATOR_PATH.with_name("phase8_ramp_summary_product_witness.py")

SOURCE_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9"
    r"\2026-07-09 ssor-prod"
)
DEFAULT_SUMMARY_PDF_ROOT = SOURCE_ROOT / "ramp_summary"
DEFAULT_SUMMARY_XLSX_ROOT = SOURCE_ROOT / "ramp_summary_excel"
DEFAULT_DETAIL_XLSX_ROOT = SOURCE_ROOT / "ramp_detail"
DEFAULT_DETAIL_PDF_ROOT = SOURCE_ROOT / "ramp_detail_pdf"
DEFAULT_TSN_RAW = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\ramp_summary\raw"
    r"\Ramp Summary Statewide_TSN.pdf"
)
R7_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd\phase4_tsn_rebaseline"
    r"\raw-2026-07-12-r7"
)
DEFAULT_TSN_XLSX = (
    R7_ROOT / "ramp_summary" / "consolidated" /
    "tsn_ramp_summary_normalized.xlsx"
)
PHASE6_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd\phase6_tsn_conservation"
)
DEFAULT_STAGE6_RESULT = PHASE6_ROOT / "ramp_summary_conservation_r7.json"
DEFAULT_STAGE6_ACCEPTANCE = Path(str(DEFAULT_STAGE6_RESULT) + ".acceptance.json")
DEFAULT_WORK_ROOT = REPO_ROOT / "tmp" / "phase8-ramp-summary-oracle"

TREE_BINDINGS = {
    "summary_pdf": {
        "files": 126,
        "bytes": 10_521_961,
        "manifest_sha256": "81108f5bb35ecffa292fd206724c2ec87001c1d0c32db33f8281a78b24f8c444",
        "prior_inventory_manifest_sha256": "dfae6c9d221e1d4929ecdb3599880e19551f92bdccd051325e498b8fbfe9d396",
        "suffix": ".pdf",
    },
    "summary_xlsx": {
        "files": 126,
        "bytes": 2_450_040,
        "manifest_sha256": "d74c19b589108e0dcbd21389f63c1adcd4d9373c4959c168a1e4ba8446c6281e",
        "prior_inventory_manifest_sha256": "450800352296dc7c8edda55190ceaa786efab6131a188428dcd2fae7c838670a",
        "suffix": ".xlsx",
    },
    "detail_xlsx": {
        "files": 126,
        "bytes": 7_858_480,
        "manifest_sha256": "7c10fbf6b996a8a9fbb0e8c8c30d8d2dac0a80c0befb7c12bdeb0151f7ff7489",
        "prior_inventory_manifest_sha256": "55a8562ddcef065b986f1d6f66ec77000081a9c75954062f51955a467d07ff51",
        "suffix": ".xlsx",
    },
    "detail_pdf": {
        "files": 126,
        "bytes": 12_792_211,
        "manifest_sha256": "6e8a2b669148738344a0173cca52a16884b972cba4679ba6446547ce8286c4c9",
        "prior_inventory_manifest_sha256": "09dff582abccbb91a58dc1cbe2b5db854469ce2b01936631bb5300dd22046aa3",
        "suffix": ".pdf",
    },
}
FILE_BINDINGS = {
    "tsn_raw_pdf": {
        "bytes": 11_758,
        "sha256": "e09842e939af4bc0da82014cfd0de1f6670e7fed5e4c5f6441628bda818a118b",
    },
    "tsn_normalized_xlsx": {
        "bytes": 5_758,
        "sha256": "15e5b9260b79618371d0378afa40f051a8912c7056c8fbf43cdbbde47b143356",
    },
    "stage6_result": {
        "bytes": 384_147,
        "sha256": "38b500489c8a310529c4c7b76bea3fe7461374d6c786b992caaa458e0ef65421",
    },
    "stage6_acceptance": {
        "bytes": 128_177,
        "sha256": "55c43d501960d3ca3702e5eac1202f96ac6c9b3e1df2eb915b19c593669bf74c",
    },
}

LEFT = (
    ("hwy_right", 8, "R - Right"),
    ("hwy_divided", 9, "D - Divided"),
    ("hwy_undivided", 10, "U - Undivided"),
    ("hwy_unconstructed", 11, "X - Unconstructed"),
    ("hwy_left", 12, "L - Left"),
    ("hwy_others", 13, "Others"),
    ("onoff_on", 17, "ON - On"),
    ("onoff_off", 18, "OFF - Off"),
    ("onoff_other", 19, "OTH - Other"),
    ("pop_rural_inside", 23, "R-RURAL -I INSIDE CITY"),
    ("pop_rural_outside", 24, "        -O OUTSIDE CITY"),
    ("pop_urban_inside", 25, "U-URBAN -I INSIDE CITY"),
    ("pop_urban_outside", 26, "        -O OUTSIDE CITY"),
    ("pop_invalid", 27, "        -INVALID DATA"),
)
RIGHT = (
    ("ramp_A_frontage", 31, "A - Frontage Road"),
    ("ramp_B_collector", 32, "B - Collector Road"),
    ("ramp_C_connector_L", 33, "C - Direct or Semi-direct Connector (Left)"),
    ("ramp_D_diamond", 34, "D - Diamond Type Ramp"),
    ("ramp_E_slip", 35, "E - Slip Ramp"),
    ("ramp_F_connector_R", 36, "F - Direct or Semi-direct Connector (Right)"),
    ("ramp_G_loop_left", 37, "G - Loop (w/Left turn)"),
    ("ramp_H_buttonhook", 38, "H - Buttonhook Ramp"),
    ("ramp_J_scissors", 39, "J - Scissors"),
    ("ramp_K_split", 40, "K - Split Ramp"),
    ("ramp_L_loop_noleft", 41, "L - Loop without Left Turn"),
    ("ramp_M_two_way", 42, "M - Two way Ramp Segment"),
    ("ramp_R_rest_area", 43, "R - Rest Area, Vista Point, Truck Scale"),
    ("ramp_Z_other", 44, "Z - Other"),
)
ALL_SOURCE = LEFT + RIGHT

COMPARE_KEYS = {
    "hwy_right": "Highway Group: R - Right",
    "hwy_divided": "Highway Group: D - Divided",
    "hwy_undivided": "Highway Group: U - Undivided",
    "hwy_unconstructed": "Highway Group: X - Unconstructed",
    "hwy_left": "Highway Group: L - Left",
    "hwy_others": "Highway Group: Others",
    "onoff_on": "On/Off: ON - On",
    "onoff_off": "On/Off: OFF - Off",
    "onoff_other": "On/Off: OTH - Other",
    "pop_rural_inside": "Population: R-RURAL -I INSIDE CITY",
    "pop_rural_outside": "Population: R-RURAL -O OUTSIDE CITY",
    "pop_urban_inside": "Population: U-URBAN -I INSIDE CITY",
    "pop_urban_outside": "Population: U-URBAN -O OUTSIDE CITY",
    "pop_invalid": "Population: -INVALID DATA",
    "ramp_A_frontage": "Ramp Type: A - Frontage Road",
    "ramp_B_collector": "Ramp Type: B - Collector Road",
    "ramp_C_connector_L": "Ramp Type: C - Direct or Semi-direct Connector (Left)",
    "ramp_D_diamond": "Ramp Type: D - Diamond Type Ramp",
    "ramp_E_slip": "Ramp Type: E - Slip Ramp",
    "ramp_F_connector_R": "Ramp Type: F - Direct or Semi-direct Connector (Right)",
    "ramp_G_loop_left": "Ramp Type: G - Loop (w/Left turn)",
    "ramp_H_buttonhook": "Ramp Type: H - Buttonhook Ramp",
    "ramp_J_scissors": "Ramp Type: J - Scissors",
    "ramp_K_split": "Ramp Type: K - Split Ramp",
    "ramp_L_loop_noleft": "Ramp Type: L - Loop without Left Turn",
    "ramp_M_two_way": "Ramp Type: M - Two way Ramp Segment",
    "ramp_R_rest_area": "Ramp Type: R - Rest Area, Vista Point, Truck Scale",
    "ramp_Z_other": "Ramp Type: Z - Other",
    "total_ramps": "Total Number of Ramps",
}
P_KEY = "Ramp Type: P - Dummy Paired"
V_KEY = "Ramp Type: V - Dummy, Volume only"
TSN_ORDER = (
    *[COMPARE_KEYS[slug] for slug, _row, _label in LEFT],
    *[COMPARE_KEYS[slug] for slug, _row, _label in RIGHT[:12]],
    P_KEY,
    COMPARE_KEYS["ramp_R_rest_area"],
    V_KEY,
    COMPARE_KEYS["ramp_Z_other"],
    COMPARE_KEYS["total_ramps"],
)
KEY_TO_SLUG = {key: slug for slug, key in COMPARE_KEYS.items()}

EXPECTED_LEFT_SIGNATURE = (
    "HighwayGroupsNUMBERCODER-RightD-DividedU-UndividedX-Unconstructed"
    "L-LeftOthersOn/OffIndicatorNUMBERCODEON-OnOFF-OffOTH-Other"
    "PopulationGroupsNUMBERCODER-RURAL-IINSIDECITY-OOUTSIDECITY"
    "U-URBAN-IINSIDECITY-OOUTSIDECITY-INVALIDDATA"
)
EXPECTED_RIGHT_SIGNATURE = (
    "RampTypesNUMBCODEERA-FrontageRoadB-CollectorRoad"
    "C-DirectorSemi-directConnectorLeftD-DiamondTypeRampE-SlipRamp"
    "F-DirectorSemi-directConnectorRightG-Loopw/LeftturnH-ButtonhookRamp"
    "J-ScissorsK-SplitRampL-LoopwithoutLeftTurnM-TwowayRampSegment"
    "R-RestAreaVistaPointTruckScaleZ-OtherTotalNumberofRamps"
    "RampPointsw/outlinework"
)
DETAIL_HEADER = (
    "Location", None, "PM", "Date of Record", None, "HG", "Area 4",
    None, "City Code", "R/U", "Description",
)
EXPECTED_PV_BY_ROUTE = {
    "005": 1, "008": 2, "010": 9, "094": 1, "110": 1,
    "134": 1, "210": 4, "280": 2, "605": 1,
}
EXPECTED_PV_BY_TYPE = {"P": 2, "V": 20}

EXPECTED_CROSS_FORMAT_SHA256 = (
    "57514b890de9d1e49ed605c0fa095fade6a264f821e8177ac19aa852d87c2f1b"
)
EXPECTED_COMPARISON_TRUTH_SHA256 = (
    "a3cbf7528aa66989f08a0d28efd8ba0e4588b8e3675ef108b0b791fdd35a2d63"
)
EXPECTED_TSMIS_AGGREGATE = {
    "hwy_right": 137,
    "hwy_divided": 14_489,
    "hwy_undivided": 85,
    "hwy_unconstructed": 0,
    "hwy_left": 171,
    "hwy_others": 334,
    "onoff_on": 7_627,
    "onoff_off": 7_405,
    "onoff_other": 125,
    "pop_rural_inside": 72,
    "pop_rural_outside": 2_437,
    "pop_urban_inside": 9_681,
    "pop_urban_outside": 3_017,
    "pop_invalid": 9,
    "ramp_A_frontage": 31,
    "ramp_B_collector": 173,
    "ramp_C_connector_L": 670,
    "ramp_D_diamond": 6_789,
    "ramp_E_slip": 334,
    "ramp_F_connector_R": 2_230,
    "ramp_G_loop_left": 615,
    "ramp_H_buttonhook": 1_154,
    "ramp_J_scissors": 294,
    "ramp_K_split": 1_002,
    "ramp_L_loop_noleft": 1_332,
    "ramp_M_two_way": 34,
    "ramp_R_rest_area": 333,
    "ramp_Z_other": 144,
    "total_ramps": 15_216,
    "ramp_points_no_linework": 59,
}

PRODUCT_DATA_COLUMNS = (
    ("Right", "hwy_right"),
    ("Divided", "hwy_divided"),
    ("Undivided", "hwy_undivided"),
    ("Unconstructed", "hwy_unconstructed"),
    ("Left", "hwy_left"),
    ("Others", "hwy_others"),
    ("ON", "onoff_on"),
    ("OFF", "onoff_off"),
    ("OTHER", "onoff_other"),
    ("Rural-Inside", "pop_rural_inside"),
    ("Rural-Outside", "pop_rural_outside"),
    ("Urban-Inside", "pop_urban_inside"),
    ("Urban-Outside", "pop_urban_outside"),
    ("Invalid", "pop_invalid"),
    ("A-Frontage", "ramp_A_frontage"),
    ("B-Collector", "ramp_B_collector"),
    ("C-Conn(L)", "ramp_C_connector_L"),
    ("D-Diamond", "ramp_D_diamond"),
    ("E-Slip", "ramp_E_slip"),
    ("F-Conn(R)", "ramp_F_connector_R"),
    ("G-LoopL", "ramp_G_loop_left"),
    ("H-Buttonhook", "ramp_H_buttonhook"),
    ("J-Scissors", "ramp_J_scissors"),
    ("K-Split", "ramp_K_split"),
    ("L-LoopNoL", "ramp_L_loop_noleft"),
    ("M-TwoWay", "ramp_M_two_way"),
    ("P-DummyPair", None),
    ("R-Rest", "ramp_R_rest_area"),
    ("V-DummyVol", None),
    ("Z-Other", "ramp_Z_other"),
    ("Total Ramps", "total_ramps"),
    ("Pts w/o Linework", "ramp_points_no_linework"),
)
PRODUCT_AUDIT_HEADERS = (
    "Sum Hwy", "Sum On/Off + NoLW", "Sum Pop",
    "Sum RampTypes + NoLW", "Audit OK",
)
PRODUCT_HEADERS = (
    "Source File", "Route",
    *[display for display, _slug in PRODUCT_DATA_COLUMNS],
    *PRODUCT_AUDIT_HEADERS,
)

FAMILIAR_DATA_ROWS = tuple(zip(
    (8, 9, 10, 11, 12, 13, 15, 16, 17, 19, 20, 21, 22, 23,
     25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38,
     39, 40, 42),
    TSN_ORDER,
))

VISUAL_REVIEW = {
    "workflow": "PDF skill: Poppler 150-dpi page-2 renders plus original/high visual inspection",
    "reviewed_utc_date": "2026-07-12",
    "sample_rationale": (
        "Route 001 first; Route 005 P-residual/dense; Route 005S suffix; Route 010 "
        "largest V-residual; Route 101 high-volume/no-linework; Route 980 last."
    ),
    "renders": [
        {"route": "001", "bytes": 111_327, "sha256": "1eeb0f94078078cefb2f45c948c5b5219cdfb6716320107254c459f09f682a86"},
        {"route": "005", "bytes": 114_390, "sha256": "e2ec270aff42c91859962398d7cb9953206bef3bd1a1f47f1c44e61414a73db9"},
        {"route": "005S", "bytes": 106_925, "sha256": "8316a0a60adffbc52f1cea0bb8c028f170d230d014a5a1e498d52565ba8e650e"},
        {"route": "010", "bytes": 112_538, "sha256": "f8df0bca4ea9e586c8bef952dc74bc220c94d0bc01a4ca465e8cc381c464dd66"},
        {"route": "101", "bytes": 115_910, "sha256": "722b2039f79c1af9c720eafefacca3eea79ec18022f2f48815fcad3dc1cd3102"},
        {"route": "980", "bytes": 108_600, "sha256": "70c7c19423cf6494786a70862637685327dd648137d9e98210e509033cc2a672"},
    ],
    "observations": (
        "Every sampled page is legible and unclipped.  The same four section blocks, "
        "14 printed Ramp Type rows, Total, and no-linework footnote are visibly present. "
        "P and V are visibly absent even on Routes 005/010, whose same-pull Detail PDFs "
        "print one P and nine V records respectively."
    ),
}

logging.getLogger("pdfminer").setLevel(logging.ERROR)


class AuditError(ValueError):
    pass


@dataclass(frozen=True)
class FileEntry:
    name: str
    bytes: int
    sha256: str


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_digest(rows: Iterable[Sequence[object]]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update((json.dumps(list(row), ensure_ascii=False,
                                  separators=(",", ":")) + "\n").encode("utf-8"))
    return digest.hexdigest()


def _strict_count(value: object, context: str) -> int:
    if type(value) is not int or value < 0:
        raise AuditError(f"{context}: expected nonnegative integer, got {value!r}")
    return value


def _route_from_name(path: Path, suffix: str) -> str:
    match = re.fullmatch(
        rf"tsar_ramp_(?:summary|detail)_route_(\d{{3}}[A-Z]?)\{suffix}",
        path.name,
    )
    if not match:
        raise AuditError(f"unexpected source filename: {path.name}")
    return match.group(1)


def _require_exact_route_universes(
        universes: dict[str, Sequence[str]], expected_count: int = 126
) -> dict[str, object]:
    if not universes:
        raise AuditError("no route universes supplied")
    canonical_label, canonical_values = next(iter(universes.items()))
    canonical = list(canonical_values)
    details = {}
    for label, raw_values in universes.items():
        values = list(raw_values)
        duplicates = sorted(
            route for route, count in Counter(values).items() if count != 1)
        if len(values) != expected_count or duplicates:
            raise AuditError(
                f"{label} route universe invalid: rows={len(values)} "
                f"duplicates={duplicates!r}")
        if values != canonical:
            missing = sorted(set(canonical) - set(values))
            extra = sorted(set(values) - set(canonical))
            raise AuditError(
                f"{label} route universe/order differs from {canonical_label}: "
                f"missing={missing!r} extra={extra!r}")
        details[label] = {
            "routes": len(values),
            "unique": len(set(values)),
            "ordered_sha256": _canonical_digest((route,) for route in values),
        }
    return {
        "canonical": canonical_label,
        "expected_count": expected_count,
        "routes": canonical,
        "universes": details,
        "all_exact": True,
    }


def _manifest(root: Path, suffix: str) -> tuple[dict[str, object], list[FileEntry]]:
    paths = sorted(root.glob(f"*{suffix}"), key=lambda path: path.name)
    entries = [FileEntry(path.name, path.stat().st_size, _sha_file(path)) for path in paths]
    serialized = "".join(
        f"{entry.name}\t{entry.bytes}\t{entry.sha256}\n" for entry in entries
    ).encode("utf-8")
    return {
        "files": len(entries),
        "bytes": sum(entry.bytes for entry in entries),
        "manifest_sha256": _sha_bytes(serialized),
        "serialization": "name\\tbytes\\tsha256\\n sorted by name",
    }, entries


def _capture_tree(label: str, root: Path, binding: dict[str, object],
                  destination: Path) -> tuple[dict[str, object], Path]:
    suffix = str(binding["suffix"])
    observed, entries = _manifest(root, suffix)
    for key in ("files", "bytes", "manifest_sha256"):
        if observed[key] != binding[key]:
            raise AuditError(f"{label} {key} drift: {observed[key]!r} != {binding[key]!r}")
    destination.mkdir(parents=True, exist_ok=False)
    for entry in entries:
        source = root / entry.name
        payload = source.read_bytes()
        if len(payload) != entry.bytes or _sha_bytes(payload) != entry.sha256:
            raise AuditError(f"{label} changed during immutable capture: {entry.name}")
        (destination / entry.name).write_bytes(payload)
    captured, _ = _manifest(destination, suffix)
    if captured != observed:
        raise AuditError(f"{label} private snapshot does not match source manifest")
    return {"binding": dict(binding), "observed": observed,
            "members": [asdict(entry) for entry in entries]}, destination


def _capture_file(label: str, source: Path, binding: dict[str, object],
                  destination: Path) -> tuple[dict[str, object], Path]:
    payload = source.read_bytes()
    observed = {"bytes": len(payload), "sha256": _sha_bytes(payload)}
    if observed != binding:
        raise AuditError(f"{label} identity drift: {observed!r} != {binding!r}")
    destination.write_bytes(payload)
    if destination.read_bytes() != payload:
        raise AuditError(f"{label} private snapshot changed after write")
    return {"binding": dict(binding), "observed": observed}, destination


def _compact_signature(text: str) -> str:
    return re.sub(r"[^A-Za-z/-]+", "", text)


def _numeric_words(page: pdfplumber.page.Page, side: str) -> list[dict[str, object]]:
    words = page.extract_words()
    if side == "left":
        selected = [word for word in words
                    if float(word["x1"]) < 90 and 40 < float(word["top"]) < 390
                    and re.fullmatch(r"[\d,]+", str(word["text"]))]
    else:
        selected = [word for word in words
                    if 300 < float(word["x0"]) < 355 and float(word["x1"]) < 355
                    and 40 < float(word["top"]) < 350
                    and re.fullmatch(r"[\d,]+", str(word["text"]))]
    return sorted(selected, key=lambda word: (float(word["top"]), float(word["x0"])))


def _parse_summary_pdfs(root: Path) -> dict[str, object]:
    records: dict[str, dict[str, int]] = {}
    geometry_rows: list[tuple[object, ...]] = []
    metadata_counts: Counter[tuple[tuple[str, str], ...]] = Counter()
    metadata_rows = []
    for path in sorted(root.glob("*.pdf"), key=lambda item: item.name):
        route = _route_from_name(path, ".pdf")
        if route in records:
            raise AuditError(f"duplicate summary PDF route identity: {route}")
        with pdfplumber.open(path) as pdf:
            if len(pdf.pages) != 2:
                raise AuditError(f"{path.name}: expected 2 pages, got {len(pdf.pages)}")
            metadata = {str(k): str(v) for k, v in (pdf.metadata or {}).items()}
            signature = tuple(sorted(metadata.items()))
            metadata_counts[signature] += 1
            required_metadata = {
                "Producer": "Skia/PDF m150",
                "Creator": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) HeadlessChrome/150.0.0.0 "
                    "Safari/537.36"),
                "Title": "TSMIS Reports",
            }
            if any(metadata.get(key) != value
                   for key, value in required_metadata.items()):
                raise AuditError(f"{path.name}: PDF producer metadata drift")
            creation = metadata.get("CreationDate")
            if (creation != metadata.get("ModDate")
                    or not re.fullmatch(r"D:20260710\d{6}\+00'00'", creation or "")):
                raise AuditError(f"{path.name}: PDF creation/modification metadata drift")
            metadata_rows.append((route, *signature))
            cover = pdf.pages[0].extract_text() or ""
            required = (
                f"All Ramps on Route {route}", f"ROUTE {route}",
                "Yunus.Shaikh@dot.ca.gov",
            )
            if not all(token in cover for token in required):
                raise AuditError(f"{path.name}: cover provenance/route claim drift")
            if not re.search(r"REPORT DATE\s*:\s*07/09/2026", cover):
                raise AuditError(f"{path.name}: report-date drift")
            if not re.search(r"REFERENCE DATE\s*:\s*07/10/2026", cover):
                raise AuditError(f"{path.name}: reference-date drift")

            page = pdf.pages[1]
            left_text = page.crop((0, 0, 300, 450)).extract_text() or ""
            right_text = page.crop((300, 0, page.width, 450)).extract_text() or ""
            if _compact_signature(left_text) != EXPECTED_LEFT_SIGNATURE:
                raise AuditError(f"{path.name}: left category label/order drift")
            if _compact_signature(right_text) != EXPECTED_RIGHT_SIGNATURE:
                raise AuditError(f"{path.name}: right category label/order drift")
            left_words = _numeric_words(page, "left")
            right_words = _numeric_words(page, "right")
            if len(left_words) != len(LEFT) or len(right_words) != len(RIGHT):
                raise AuditError(
                    f"{path.name}: category count bands left={len(left_words)} "
                    f"right={len(right_words)}")
            record = {
                slug: int(str(word["text"]).replace(",", ""))
                for (slug, _row, _label), word in zip(LEFT, left_words)
            }
            record.update({
                slug: int(str(word["text"]).replace(",", ""))
                for (slug, _row, _label), word in zip(RIGHT, right_words)
            })
            page_text = page.extract_text() or ""
            total = re.search(r"Total Number of Ramps:\s*([\d,]+)", page_text, re.I)
            no_linework = re.search(
                r"Ramp Points w/out linework:\s*([\d,]+)", page_text, re.I)
            if not total or not no_linework:
                raise AuditError(f"{path.name}: total/footer missing")
            record["total_ramps"] = int(total.group(1).replace(",", ""))
            record["ramp_points_no_linework"] = int(no_linework.group(1).replace(",", ""))
            records[route] = record
            geometry_rows.append((
                route,
                *[round(float(word["top"]), 6) for word in left_words],
                *[round(float(word["top"]), 6) for word in right_words],
            ))
    if len(records) != 126 or len(metadata_counts) != 95:
        raise AuditError(
            f"summary PDF universe/metadata drift: routes={len(records)} "
            f"metadata_signatures={len(metadata_counts)}")
    return {
        "records": records,
        "routes": list(records),
        "pages": len(records) * 2,
        "metadata_signatures": [
            {"count": count, "metadata": dict(signature)}
            for signature, count in metadata_counts.items()
        ],
        "metadata_signature_count": len(metadata_counts),
        "metadata_ordered_typed_sha256": _canonical_digest(metadata_rows),
        "geometry_typed_sha256": _canonical_digest(geometry_rows),
    }


def _parse_summary_xlsx(root: Path) -> dict[str, object]:
    records: dict[str, dict[str, int]] = {}
    timestamps = []
    formula_cells = []
    for path in sorted(root.glob("*.xlsx"), key=lambda item: item.name):
        route = _route_from_name(path, ".xlsx")
        if route in records:
            raise AuditError(f"duplicate summary XLSX route identity: {route}")
        wb = load_workbook(path, read_only=False, data_only=False)
        try:
            if wb.sheetnames != ["Ramp Summary"]:
                raise AuditError(f"{path.name}: sheet universe {wb.sheetnames!r}")
            ws = wb["Ramp Summary"]
            if ws.max_row != 47 or ws.max_column != 2:
                raise AuditError(f"{path.name}: dimensions {ws.max_row}x{ws.max_column}")
            fixed_top = (
                ws["A1"].value == "TSAR - RAMPS SUMMARY",
                ws["A2"].value == f"All Ramps on Route {route}",
                ws["A3"].value == "Reference Date: 2026-07-10",
            )
            if not all(fixed_top):
                raise AuditError(f"{path.name}: title/route/reference-date drift")
            stamp = str(ws["A4"].value)
            if not re.fullmatch(r"Generated: 7/9/2026, \d{1,2}:\d{2}:\d{2} [AP]M", stamp):
                raise AuditError(f"{path.name}: generated timestamp drift: {stamp!r}")
            timestamps.append((route, stamp))
            fixed = {
                5: (None, None), 6: ("Highway Groups", None), 7: ("NUMBER", "CODE"),
                14: (None, None), 15: ("On/Off Indicator", None), 16: ("NUMBER", "CODE"),
                20: (None, None), 21: ("Population Groups", None), 22: ("NUMBER", "CODE"),
                28: (None, None), 29: ("Ramp Types", None), 30: ("NUMBER", "CODE"),
                45: (None, None),
            }
            for row_number, expected in fixed.items():
                actual = (ws.cell(row_number, 1).value, ws.cell(row_number, 2).value)
                if actual != expected:
                    raise AuditError(
                        f"{path.name}: row {row_number} {actual!r} != {expected!r}")
            record = {}
            for slug, row_number, label in ALL_SOURCE:
                if ws.cell(row_number, 2).value != label:
                    raise AuditError(f"{path.name}: row {row_number} label drift")
                record[slug] = _strict_count(
                    ws.cell(row_number, 1).value, f"{path.name} row {row_number}")
            if (ws["A46"].value, ws["A47"].value) != (
                    "Total Number of Ramps:", "Ramp Points w/out linework:"):
                raise AuditError(f"{path.name}: footer label drift")
            record["total_ramps"] = _strict_count(ws["B46"].value, f"{path.name} B46")
            record["ramp_points_no_linework"] = _strict_count(
                ws["B47"].value, f"{path.name} B47")
            formula_cells.extend(
                (path.name, cell.coordinate)
                for row in ws.iter_rows() for cell in row if cell.data_type == "f")
            records[route] = record
        finally:
            wb.close()
    if len(records) != 126 or formula_cells:
        raise AuditError(
            f"summary XLSX universe/formulas drift: routes={len(records)} "
            f"formulas={formula_cells[:5]!r}")
    return {"records": records, "routes": list(records),
            "generated_timestamps": timestamps}


def _parse_detail_xlsx(root: Path) -> dict[str, object]:
    route_counts: dict[str, int] = {}
    typed_rows = []
    formulas = []
    errors = []
    location_re = re.compile(r"^\d{2}-[A-Z]{2,3}-(\d{3}[A-Z]?)$")
    for path in sorted(root.glob("*.xlsx"), key=lambda item: item.name):
        route = _route_from_name(path, ".xlsx")
        if route in route_counts:
            raise AuditError(f"duplicate detail XLSX route identity: {route}")
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            if wb.sheetnames != ["TSAR - Ramp Detail"]:
                raise AuditError(f"{path.name}: detail sheet universe {wb.sheetnames!r}")
            ws = wb["TSAR - Ramp Detail"]
            rows = ws.iter_rows()
            header_cells = next(rows, ())
            header = tuple(cell.value for cell in header_cells)
            if header != DETAIL_HEADER or ws.max_column != len(DETAIL_HEADER):
                raise AuditError(
                    f"{path.name}: detail header/dimension drift: {header!r}")
            count = 0
            for source_row, cells in enumerate(rows, 2):
                values = tuple(cell.value for cell in cells)
                if all(value is None for value in values):
                    raise AuditError(f"{path.name}: blank physical detail row {source_row}")
                match = location_re.fullmatch(str(values[0] or ""))
                if not match or match.group(1) != route:
                    raise AuditError(
                        f"{path.name}: row {source_row} location/filename mismatch "
                        f"{values[0]!r}")
                for cell in cells:
                    if cell.data_type == "f":
                        formulas.append((path.name, cell.coordinate))
                    if cell.data_type == "e":
                        errors.append((path.name, cell.coordinate, cell.value))
                typed_rows.append((route, source_row, *values))
                count += 1
            route_counts[route] = count
        finally:
            wb.close()
    if len(route_counts) != 126 or formulas or errors:
        raise AuditError(
            f"detail XLSX universe/formula/error drift: routes={len(route_counts)} "
            f"formulas={formulas[:3]!r} errors={errors[:3]!r}")
    return {
        "routes": list(route_counts),
        "route_counts": route_counts,
        "rows": sum(route_counts.values()),
        "ordered_typed_sha256": _canonical_digest(typed_rows),
    }


DETAIL_LINE_RE = re.compile(r"^\d{2}-[A-Z]{2,3}-(\d{3}[A-Z]?)\s")
DETAIL_PV_RE = re.compile(
    r"^(?P<location>\d{2}-[A-Z]{2,3}-(?P<route>\d{3}[A-Z]?))\s+"
    r"(?:(?P<prefix>[RL])\s+)?(?P<pm>\d{3}\.\d{3})\s+"
    r"(?P<date>\d{2}/\d{2}/\d{4})\s+(?P<hg>[RDUXL])\s+"
    r"(?P<area4>[YN-])\s+(?:(?P<city>[A-Z0-9-]+)\s+)?"
    r"(?P<ru>[RU])\s+(?P<onoff>[NF-])\s+(?P<ramp_type>[PV])\s+"
    r"(?P<description>.*)$"
)


def _parse_detail_pdfs(root: Path) -> dict[str, object]:
    route_counts: dict[str, int] = {}
    page_counts: dict[str, int] = {}
    physical_rows = []
    pv_records = []
    for path in sorted(root.glob("*.pdf"), key=lambda item: item.name):
        route = _route_from_name(path, ".pdf")
        if route in route_counts:
            raise AuditError(f"duplicate detail PDF route identity: {route}")
        reader = PdfReader(str(path))
        if len(reader.pages) < 2:
            raise AuditError(f"{path.name}: detail PDF has no data page")
        cover = reader.pages[0].extract_text() or ""
        if f"REPORT TITLE :Route {route}" not in cover or f"ROUTE {route}" not in cover:
            raise AuditError(f"{path.name}: detail cover route drift")
        count = 0
        for page_number, page in enumerate(reader.pages[1:], 2):
            for line in (page.extract_text() or "").splitlines():
                match = DETAIL_LINE_RE.match(line)
                if not match:
                    continue
                if match.group(1) != route:
                    raise AuditError(
                        f"{path.name} page {page_number}: row route {match.group(1)!r}")
                count += 1
                physical_rows.append((route, page_number, line))
                pv = DETAIL_PV_RE.match(line)
                if pv:
                    record = {"file_route": route, "page": page_number,
                              **pv.groupdict()}
                    pv_records.append(record)
        route_counts[route] = count
        page_counts[route] = len(reader.pages)
    pv_by_route = dict(sorted(Counter(record["file_route"] for record in pv_records).items()))
    pv_by_type = dict(sorted(Counter(record["ramp_type"] for record in pv_records).items()))
    if len(route_counts) != 126:
        raise AuditError(f"detail PDF route universe drift: {len(route_counts)}")
    return {
        "routes": list(route_counts),
        "route_counts": route_counts,
        "rows": sum(route_counts.values()),
        "pages": sum(page_counts.values()),
        "data_pages": sum(count - 1 for count in page_counts.values()),
        "ordered_physical_row_sha256": _canonical_digest(physical_rows),
        "pv_records": pv_records,
        "pv_ordered_sha256": _canonical_digest(
            (record["file_route"], record["page"], record["location"],
             record["prefix"], record["pm"], record["date"], record["hg"],
             record["area4"], record["city"], record["ru"], record["onoff"],
             record["ramp_type"], record["description"])
            for record in pv_records),
        "pv_by_route": pv_by_route,
        "pv_by_type": pv_by_type,
    }


def _load_tsn_normalized(path: Path) -> dict[str, object]:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        if wb.sheetnames != ["Ramp Summary (TSN)"]:
            raise AuditError(f"TSN sheet universe {wb.sheetnames!r}")
        rows = list(wb["Ramp Summary (TSN)"].iter_rows(values_only=True))
        if not rows or rows[0] != ("Category", "Count") or len(rows) != 32:
            raise AuditError("TSN normalized header/row-count drift")
        out = {}
        order = []
        for source_row, row in enumerate(rows[1:], 2):
            key, raw_count = row
            if not isinstance(key, str) or key in out:
                raise AuditError(f"TSN row {source_row}: invalid/duplicate key {key!r}")
            count = _strict_count(raw_count, f"TSN row {source_row}")
            out[key] = count
            order.append(key)
        if tuple(order) != TSN_ORDER:
            raise AuditError("TSN normalized category order/universe drift")
        return {
            "order": order,
            "counts": out,
            "ordered_typed_sha256": _canonical_digest(
                (key, out[key]) for key in order),
        }
    finally:
        wb.close()


def _load_stage6(result_path: Path, acceptance_path: Path) -> dict[str, object]:
    result = json.loads(result_path.read_text(encoding="utf-8"))
    acceptance = json.loads(acceptance_path.read_text(encoding="utf-8"))
    checks = {
        "result_audit_complete": result.get("stage6_family_audit_complete") is True,
        "result_projection_exact": result.get("projection_exact") is True,
        "acceptance_explicit_true": acceptance.get("accepted") is True,
        "acceptance_audit_complete": acceptance.get("stage6_family_audit_complete") is True,
        "acceptance_projection_exact": acceptance.get("projection_exact") is True,
        "acceptance_result_bytes_exact": acceptance.get("result_bytes") == FILE_BINDINGS[
            "stage6_result"]["bytes"],
        "acceptance_result_sha_exact": acceptance.get("result_sha256") == FILE_BINDINGS[
            "stage6_result"]["sha256"],
    }
    return {"checks": checks,
            "normalized_full_conservation": result.get("normalized_full_conservation"),
            "accepted_open_findings": acceptance.get("open_product_findings_authorized")}


def _cross_format(pdf_records: dict[str, dict[str, int]],
                  xlsx_records: dict[str, dict[str, int]]) -> dict[str, object]:
    if list(pdf_records) != list(xlsx_records):
        raise AuditError("summary PDF/XLSX route order or universe differs")
    rows = []
    differences = []
    for route in pdf_records:
        if tuple(pdf_records[route]) != tuple(xlsx_records[route]):
            raise AuditError(f"summary field universe differs on Route {route}")
        for field in pdf_records[route]:
            left = pdf_records[route][field]
            right = xlsx_records[route][field]
            rows.append((route, field, left, right))
            if left != right:
                differences.append({"route": route, "field": field,
                                    "pdf": left, "xlsx": right})
    return {
        "routes": len(pdf_records),
        "compared_values": len(rows),
        "difference_count": len(differences),
        "differences": differences,
        "ordered_typed_sha256": _canonical_digest(rows),
    }


def _arithmetic(records: dict[str, dict[str, int]],
                pv_by_route: dict[str, int]) -> dict[str, object]:
    rows = []
    for route, record in records.items():
        total = record["total_ramps"]
        no_linework = record["ramp_points_no_linework"]
        hwy = sum(record[slug] for slug, _row, _label in LEFT[:6])
        onoff = sum(record[slug] for slug, _row, _label in LEFT[6:9])
        population = sum(record[slug] for slug, _row, _label in LEFT[9:14])
        printed_types = sum(record[slug] for slug, _row, _label in RIGHT)
        residual = total - printed_types - no_linework
        pv = pv_by_route.get(route, 0)
        rows.append({
            "route": route,
            "total": total,
            "no_linework": no_linework,
            "highway_delta": hwy - total,
            "onoff_plus_no_linework_delta": onoff + no_linework - total,
            "population_delta": population - total,
            "unprinted_ramp_type_residual": residual,
            "same_pull_detail_pv_count": pv,
            "residual_after_detail_pv": residual - pv,
        })
    return {
        "routes": rows,
        "nonzero_highway": [row for row in rows if row["highway_delta"]],
        "nonzero_onoff": [row for row in rows if row["onoff_plus_no_linework_delta"]],
        "nonzero_population": [row for row in rows if row["population_delta"]],
        "unprinted_ramp_type_routes": [
            row for row in rows if row["unprinted_ramp_type_residual"]],
        "unexplained_after_detail_pv": [
            row for row in rows if row["residual_after_detail_pv"]],
    }


def _aggregate(records: dict[str, dict[str, int]]) -> dict[str, int]:
    first = next(iter(records.values()))
    return {field: sum(record[field] for record in records.values()) for field in first}


def _comparison_truth(tsmis_aggregate: dict[str, int],
                      tsn_counts: dict[str, int]) -> dict[str, object]:
    rows = []
    for key in TSN_ORDER:
        slug = KEY_TO_SLUG.get(key)
        present = slug is not None
        tsmis = tsmis_aggregate[slug] if slug else None
        tsn = tsn_counts[key]
        rows.append({
            "category": key,
            "status": "Both" if present else "Only in TSN",
            "tsmis_present": present,
            "tsmis": tsmis,
            "tsn": tsn,
            "delta_tsn_minus_tsmis": tsn - tsmis if present else None,
            "delta_tsmis_minus_tsn": tsmis - tsn if present else None,
            "paired_equal": bool(present and tsmis == tsn),
        })
    return {
        "rows": rows,
        "both": sum(row["status"] == "Both" for row in rows),
        "only_in_tsn": sum(row["status"] == "Only in TSN" for row in rows),
        "only_in_tsmis": 0,
        "identical_both": sum(row["paired_equal"] for row in rows),
        "differing_both": sum(
            row["status"] == "Both" and not row["paired_equal"] for row in rows),
        "footnote": {
            "category": "Ramp Points w/out linework",
            "tsmis": tsmis_aggregate["ramp_points_no_linework"],
            "disposition": "display_only_metadata_excluded_from_comparison_universe",
        },
        "ordered_typed_sha256": _canonical_digest(
            (row["category"], row["status"], row["tsmis"], row["tsn"],
             row["delta_tsmis_minus_tsn"])
            for row in rows),
        "digest_delta_direction": "TSMIS minus TSN",
    }


def _typed_cell_value(value: object) -> dict[str, object]:
    if value is None:
        return {"type": "none", "value": None}
    if isinstance(value, bool):
        return {"type": "bool", "value": value}
    if isinstance(value, int):
        return {"type": "int", "value": value}
    if isinstance(value, float):
        return {"type": "float", "value": repr(value)}
    if isinstance(value, (datetime, date, time)):
        return {"type": type(value).__name__, "value": value.isoformat()}
    if isinstance(value, bytes):
        return {"type": "bytes", "value": value.hex()}
    return {"type": type(value).__name__, "value": str(value)}


def _zip_digest_without_core(path: Path) -> dict[str, object]:
    rows = []
    excluded = []
    with zipfile.ZipFile(path) as archive:
        names = [info.filename for info in archive.infolist()]
        duplicates = sorted(name for name, count in Counter(names).items() if count != 1)
        if duplicates:
            raise AuditError(f"{path.name}: duplicate ZIP members {duplicates!r}")
        bad = archive.testzip()
        if bad is not None:
            raise AuditError(f"{path.name}: corrupt ZIP member {bad!r}")
        for name in sorted(names):
            payload = archive.read(name)
            if name == "docProps/core.xml":
                # Deliberately retain only the excluded role, never its volatile
                # timestamp-bearing bytes/hash.  The stable replay record must not
                # smuggle the ignored nondeterminism back into its own equality.
                excluded.append(name)
                continue
            rows.append((name, len(payload), _sha_bytes(payload)))
    return {
        "member_count": len(names),
        "included_member_count": len(rows),
        "included_uncompressed_bytes": sum(row[1] for row in rows),
        "excluded_members": excluded,
        "canonical_member_sha256": _canonical_digest(rows),
    }


def _workbook_semantic_digest(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    rows: list[tuple[object, ...]] = []
    cell_count = 0
    formula_count = 0
    try:
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            ws = workbook[sheet_name]
            rows.append(("sheet", sheet_index, sheet_name, ws.sheet_state,
                         ws.max_row, ws.max_column, str(ws.freeze_panes or ""),
                         str(ws.auto_filter.ref or "")))
            for merged in sorted(str(item) for item in ws.merged_cells.ranges):
                rows.append(("merge", sheet_name, merged))
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    typed = _typed_cell_value(cell.value)
                    rows.append(("cell", sheet_name, cell.coordinate, cell.data_type,
                                 typed["type"], typed["value"], cell.number_format))
                    cell_count += 1
                    formula_count += cell.data_type == "f"
    finally:
        workbook.close()
    return {
        "sheets": len([row for row in rows if row[0] == "sheet"]),
        "nonblank_cells": cell_count,
        "formula_cells": formula_count,
        "ordered_semantic_sha256": _canonical_digest(rows),
    }


def _product_int(value: object, context: str) -> int:
    if isinstance(value, bool):
        raise AuditError(f"{context}: Boolean is not a count")
    if isinstance(value, int) and value >= 0:
        return value
    if isinstance(value, str) and re.fullmatch(r"0|[1-9]\d*", value):
        return int(value)
    raise AuditError(f"{context}: expected product integer/literal, got {value!r}")


def _workbook_strings(workbook: openpyxl.Workbook) -> list[str]:
    return [
        cell.value
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]


def _expected_product_formulas(row_number: int) -> dict[str, str]:
    return {
        "Sum Hwy": f"=SUM(C{row_number}:H{row_number})",
        "Sum On/Off + NoLW": f"=SUM(I{row_number}:K{row_number})+AH{row_number}",
        "Sum Pop": f"=SUM(L{row_number}:P{row_number})",
        "Sum RampTypes + NoLW": f"=SUM(Q{row_number}:AF{row_number})+AH{row_number}",
        "Audit OK": (
            f'=IF(AND(AI{row_number}=AG{row_number},AJ{row_number}=AG{row_number},'
            f'AK{row_number}=AG{row_number},AL{row_number}=AG{row_number}),"OK",'
            f'"⚠ Source ≠ total: "&IF(AI{row_number}<>AG{row_number},"Hwy ","")&'
            f'IF(AJ{row_number}<>AG{row_number},"On/Off ","")&'
            f'IF(AK{row_number}<>AG{row_number},"Pop ","")&'
            f'IF(AL{row_number}<>AG{row_number},"Ramp Types ",""))'
        ),
    }


def _inspect_production_consolidated(
        path: Path, raw_records: dict[str, dict[str, int]]) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    mismatches = []
    synthetic_nonblank = []
    formula_mismatches = []
    found_routes = []
    try:
        if workbook.sheetnames != ["Combined", "TSAR Ramps Summary"]:
            raise AuditError(
                f"production consolidated sheet universe {workbook.sheetnames!r}")
        ws = workbook["TSAR Ramps Summary"]
        if (ws.max_row, ws.max_column) != (128, 39):
            raise AuditError(
                f"production summary dimensions {ws.max_row}x{ws.max_column}")
        headers = tuple(ws.cell(2, column).value for column in range(1, 40))
        if headers != PRODUCT_HEADERS:
            raise AuditError("production consolidated column schema/order drift")
        positions = {header: index + 1 for index, header in enumerate(headers)}
        for row_number in range(3, 129):
            route = ws.cell(row_number, positions["Route"]).value
            if not isinstance(route, str) or route not in raw_records:
                raise AuditError(
                    f"production row {row_number}: invalid route {route!r}")
            if route in found_routes:
                raise AuditError(f"production duplicate route {route!r}")
            found_routes.append(route)
            expected_file = f"tsar_ramp_summary_route_{route}.pdf"
            source_file = ws.cell(row_number, positions["Source File"]).value
            if source_file != expected_file:
                raise AuditError(
                    f"production row {row_number}: source {source_file!r} != "
                    f"{expected_file!r}")
            for display, slug in PRODUCT_DATA_COLUMNS:
                cell = ws.cell(row_number, positions[display])
                if cell.data_type == "f":
                    raise AuditError(
                        f"production source value became formula: {cell.coordinate}")
                if slug is None:
                    if cell.value is not None:
                        synthetic_nonblank.append(
                            {"route": route, "column": display, "value": cell.value})
                    continue
                expected = raw_records[route][slug]
                if cell.value != expected or type(cell.value) is not int:
                    mismatches.append({
                        "route": route, "field": slug, "expected": expected,
                        "observed": cell.value, "type": type(cell.value).__name__,
                    })
            for display, expected_formula in _expected_product_formulas(
                    row_number).items():
                cell = ws.cell(row_number, positions[display])
                if cell.data_type != "f" or cell.value != expected_formula:
                    formula_mismatches.append({
                        "route": route, "cell": cell.coordinate,
                        "expected": expected_formula, "observed": cell.value,
                    })

        if found_routes != list(raw_records):
            raise AuditError("production consolidated route order/universe drift")

        combined = workbook["Combined"]
        if (combined.max_row, combined.max_column) != (29, 8):
            raise AuditError(
                f"production Combined dimensions {combined.max_row}x{combined.max_column}")
        combined_targets = {
            **{f"A{6 + index}": slug for index, (slug, _row, _label) in enumerate(LEFT[:6])},
            **{f"A{15 + index}": slug for index, (slug, _row, _label) in enumerate(LEFT[6:9])},
            **{f"A{21 + index}": slug for index, (slug, _row, _label) in enumerate(LEFT[9:14])},
            **{f"E{6 + index}": slug for index, slug in enumerate((
                "ramp_A_frontage", "ramp_B_collector", "ramp_C_connector_L",
                "ramp_D_diamond", "ramp_E_slip", "ramp_F_connector_R",
                "ramp_G_loop_left", "ramp_H_buttonhook", "ramp_J_scissors",
                "ramp_K_split", "ramp_L_loop_noleft", "ramp_M_two_way",
                None, "ramp_R_rest_area", None, "ramp_Z_other"))},
            "C28": "total_ramps",
            "C29": "ramp_points_no_linework",
        }
        slug_to_display = {slug: display for display, slug in PRODUCT_DATA_COLUMNS
                           if slug is not None}
        synthetic_display = {"E18": "P-DummyPair", "E20": "V-DummyVol"}
        combined_formula_mismatches = []
        for coordinate, slug in combined_targets.items():
            display = synthetic_display.get(coordinate) if slug is None else slug_to_display[slug]
            source_column = get_column_letter(positions[display])
            expected_formula = (
                f"=SUM('TSAR Ramps Summary'!{source_column}3:{source_column}128)")
            cell = combined[coordinate]
            if cell.data_type != "f" or cell.value != expected_formula:
                combined_formula_mismatches.append({
                    "cell": coordinate, "expected": expected_formula,
                    "observed": cell.value,
                })

        strings = _workbook_strings(workbook)
        provenance_tokens = (
            "07/09/2026", "07/10/2026", "Yunus.Shaikh@dot.ca.gov",
            "Generated:", "Reference Date:", "REPORT DATE", "REFERENCE DATE",
        )
        provenance_presence = {
            token: any(token in value for value in strings)
            for token in provenance_tokens
        }
        cf_ranges = sorted(str(item.sqref) for item in ws.conditional_formatting)
    finally:
        workbook.close()

    arithmetic = _arithmetic(raw_records, EXPECTED_PV_BY_ROUTE)
    warning_routes = [
        row["route"] for row in arithmetic["unprinted_ramp_type_routes"]]
    if mismatches or synthetic_nonblank or formula_mismatches or combined_formula_mismatches:
        raise AuditError(
            "production consolidated projection/formula drift: "
            f"values={len(mismatches)} synthetic={len(synthetic_nonblank)} "
            f"row_formulas={len(formula_mismatches)} "
            f"combined={len(combined_formula_mismatches)}")
    if warning_routes != list(EXPECTED_PV_BY_ROUTE):
        raise AuditError(f"production audit-warning route set drift: {warning_routes!r}")
    if cf_ranges != ["AM3:AM128"]:
        raise AuditError(f"production audit conditional-format range drift: {cf_ranges!r}")
    return {
        "package_without_core": _zip_digest_without_core(path),
        "semantic_digest": _workbook_semantic_digest(path),
        "routes": len(found_routes),
        "source_backed_values_compared": len(found_routes) * 30,
        "source_backed_value_mismatches": 0,
        "synthetic_pv_cells": len(found_routes) * 2,
        "synthetic_pv_nonblank": 0,
        "row_audit_formula_cells": len(found_routes) * 5,
        "combined_formula_cells": len(combined_targets),
        "audit_warning_routes": warning_routes,
        "audit_warning_route_count": len(warning_routes),
        "conditional_format_ranges": cf_ranges,
        "printed_tsmis_provenance_presence": provenance_presence,
        "printed_tsmis_provenance_preserved": any(provenance_presence.values()),
        "projection_exact": True,
    }


EXPECTED_PRODUCT_GAP_IDS = (
    "ramp_type_p_fabricated_tsmis_zero",
    "ramp_type_v_fabricated_tsmis_zero",
    "no_linework_display_metric_injected_into_comparison",
)


def _semantic_gaps(product_rows: Sequence[dict[str, object]],
                   truth_rows: Sequence[dict[str, object]]) -> list[dict[str, object]]:
    truth_by_key = {str(row["category"]): row for row in truth_rows}
    product_by_key = {str(row["category"]): row for row in product_rows}
    if len(product_by_key) != len(product_rows):
        return [{"id": "duplicate_product_comparison_category"}]
    gaps = []
    for key in (P_KEY, V_KEY):
        expected = truth_by_key[key]
        observed = product_by_key.get(key)
        expected_id = (
            "ramp_type_p_fabricated_tsmis_zero" if key == P_KEY
            else "ramp_type_v_fabricated_tsmis_zero")
        if observed is None:
            gaps.append({"id": f"missing_{expected_id}", "category": key})
        elif (observed.get("status") != expected["status"]
              or observed.get("tsmis") != expected["tsmis"]):
            gaps.append({
                "id": expected_id,
                "category": key,
                "intended": {"status": expected["status"], "tsmis": expected["tsmis"]},
                "product": {"status": observed.get("status"),
                            "tsmis": observed.get("tsmis")},
            })
    for key, expected in truth_by_key.items():
        if key in (P_KEY, V_KEY):
            continue
        observed = product_by_key.get(key)
        if observed is None:
            gaps.append({"id": "missing_comparison_category", "category": key})
            continue
        if (observed.get("status") != expected["status"]
                or observed.get("tsmis") != expected["tsmis"]
                or observed.get("tsn") != expected["tsn"]):
            gaps.append({
                "id": "source_backed_comparison_row_mismatch",
                "category": key, "intended": expected, "product": observed,
            })
    for row in product_rows:
        key = str(row["category"])
        if key in truth_by_key:
            continue
        if key == "Ramp Points w/out linework":
            gaps.append({
                "id": "no_linework_display_metric_injected_into_comparison",
                "category": key,
                "intended": "display-only metadata excluded from comparison universe",
                "product": {"status": row.get("status"), "tsmis": row.get("tsmis"),
                            "tsn": row.get("tsn")},
            })
        else:
            gaps.append({"id": "unexpected_product_comparison_category",
                         "category": key})
    return gaps


def _inspect_production_comparison(
        formulas_path: Path, values_path: Path, truth: dict[str, object],
        tsmis_aggregate: dict[str, int], tsn_counts: dict[str, int],
        helper: dict[str, object]) -> dict[str, object]:
    expected_sheets = [
        "Summary", "Spot Check", "Comparison", "Only in TSMIS", "Only in TSN",
        "TSMIS", "TSN", "Summary by Category",
        "__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B",
    ]
    formulas = load_workbook(formulas_path, read_only=False, data_only=False)
    values = load_workbook(values_path, read_only=False, data_only=False)
    try:
        if formulas.sheetnames != expected_sheets or values.sheetnames != expected_sheets:
            raise AuditError(
                "production comparison sheet universe drift: "
                f"formulas={formulas.sheetnames!r} values={values.sheetnames!r}")
        if not any(
                cell.data_type == "f"
                for ws in formulas.worksheets for row in ws.iter_rows() for cell in row):
            raise AuditError("production formulas flavor contains no formulas")

        tsmis_ws = values["TSMIS"]
        tsn_ws = values["TSN"]
        if tuple(tsmis_ws.cell(1, col).value for col in range(1, 6)) != (
                "Comparison row", "Category", "Count", "Key (helper)",
                "__CMP_E2_BUILD_FRESH_V1_C001_B_D"):
            raise AuditError("production comparison TSMIS header drift")
        if tuple(tsn_ws.cell(1, col).value for col in range(1, 6)) != (
                "Comparison row", "Category", "Count", "Key (helper)",
                "__CMP_E2_BUILD_FRESH_V1_C001_B_D"):
            raise AuditError("production comparison TSN header drift")

        tsmis_order = [tsmis_ws.cell(row, 2).value for row in range(2, 34)]
        tsn_order = [tsn_ws.cell(row, 2).value for row in range(2, 33)]
        expected_tsmis_order = [*TSN_ORDER, "Ramp Points w/out linework"]
        if tsmis_order != expected_tsmis_order or tsn_order != list(TSN_ORDER):
            raise AuditError("production comparison input category order/universe drift")
        product_tsmis = {
            str(tsmis_ws.cell(row, 2).value): _product_int(
                tsmis_ws.cell(row, 3).value, f"TSMIS comparison row {row}")
            for row in range(2, 34)
        }
        product_tsn = {
            str(tsn_ws.cell(row, 2).value): _product_int(
                tsn_ws.cell(row, 3).value, f"TSN comparison row {row}")
            for row in range(2, 33)
        }
        for key in TSN_ORDER:
            if product_tsn[key] != tsn_counts[key]:
                raise AuditError(f"production TSN value drift for {key!r}")
            slug = KEY_TO_SLUG.get(key)
            expected_tsmis = tsmis_aggregate[slug] if slug else 0
            if product_tsmis[key] != expected_tsmis:
                raise AuditError(f"production TSMIS value drift for {key!r}")
        if product_tsmis["Ramp Points w/out linework"] != tsmis_aggregate[
                "ramp_points_no_linework"]:
            raise AuditError("production no-linework value drift")

        comparison = values["Comparison"]
        if (comparison.max_row, comparison.max_column) != (33, 8):
            raise AuditError(
                f"production Comparison dimensions {comparison.max_row}x"
                f"{comparison.max_column}")
        expected_comparison_header = (
            "Category", "#", "TSMIS Row", "TSN Row", "Status", "Diffs", "Count",
            "__CMP_E1_STATE_V1_C001_P0000_P0000",
        )
        if tuple(comparison.cell(1, col).value for col in range(1, 9)) != (
                expected_comparison_header):
            raise AuditError("production Comparison header drift")
        product_rows = []
        expected_product_order = [*TSN_ORDER, "Ramp Points w/out linework"]
        for row_number, key in enumerate(expected_product_order, 2):
            if comparison.cell(row_number, 1).value != key:
                raise AuditError(
                    f"production Comparison row {row_number} category drift")
            status = comparison.cell(row_number, 5).value
            diffs = comparison.cell(row_number, 6).value
            state = comparison.cell(row_number, 8).value
            left = product_tsmis.get(key)
            right = product_tsn.get(key)
            expected_status = "Both" if right is not None else "TSMIS only"
            expected_diffs = (int(left != right) if right is not None else None)
            expected_state = "D" if expected_diffs else "E" if right is not None else "U"
            if (status, diffs, state) != (
                    expected_status, expected_diffs, expected_state):
                raise AuditError(
                    f"production Comparison row {row_number} state drift: "
                    f"{(status, diffs, state)!r}")
            product_rows.append({
                "category": key,
                "status": status,
                "tsmis": left,
                "tsn": right,
                "diffs": diffs,
                "state": state,
            })

        only_tsmis = values["Only in TSMIS"]
        only_tsn = values["Only in TSN"]
        if (only_tsmis.max_row, only_tsmis.max_column) != (2, 4):
            raise AuditError("production Only in TSMIS shape drift")
        if only_tsmis["A2"].value != "Ramp Points w/out linework":
            raise AuditError("production TSMIS-only category drift")
        if (only_tsn.max_row, only_tsn.max_column) != (1, 4):
            raise AuditError("production Only in TSN shape drift")

        familiar = values["Summary by Category"]
        if (familiar.max_row, familiar.max_column) != (45, 4):
            raise AuditError(
                f"production familiar sheet dimensions {familiar.max_row}x"
                f"{familiar.max_column}")
        familiar_mismatches = []
        truth_by_key = {
            str(row["category"]): row for row in truth["rows"]}
        for row_number, key in FAMILIAR_DATA_ROWS:
            intended = truth_by_key[key]
            left = 0 if key in (P_KEY, V_KEY) else intended["tsmis"]
            right = intended["tsn"]
            observed = tuple(familiar.cell(row_number, col).value for col in range(2, 5))
            expected = (left, right, right - left)
            if observed != expected:
                familiar_mismatches.append({
                    "row": row_number, "category": key,
                    "expected": expected, "observed": observed,
                })
        if tuple(familiar.cell(45, col).value for col in range(2, 5)) != (
                tsmis_aggregate["ramp_points_no_linework"], None, None):
            familiar_mismatches.append({"row": 45, "category": "no_linework"})
        if familiar_mismatches:
            raise AuditError(
                f"production familiar numeric projection drift: {familiar_mismatches[:3]!r}")

        note_row_2 = str(familiar["A2"].value or "")
        note_row_4 = str(familiar["A4"].value or "")
        note_claims = {
            "zero_fills_unclassified_categories": (
                "show 0 on that side" in note_row_2),
            "pv_stay_one_sided_by_design": (
                "stay one-sided by design" in note_row_4),
            "footnote_never_compared": "never compared" in note_row_4,
        }
        if not all(note_claims.values()):
            raise AuditError(f"production familiar note contract drift: {note_claims!r}")

        all_strings = _workbook_strings(values)
        printed_tokens = (
            "07/09/2026", "07/10/2026", "Yunus.Shaikh@dot.ca.gov",
            "REPORT DATE", "REFERENCE DATE", "Generated:",
        )
        printed_presence = {
            token: any(token in value for value in all_strings)
            for token in printed_tokens
        }
        source_note = str(values["Summary"]["B5"].value or "")
        familiar_source_note = str(familiar["A3"].value or "")
    finally:
        formulas.close()
        values.close()

    gaps = _semantic_gaps(product_rows, truth["rows"])
    gap_ids = tuple(gap["id"] for gap in gaps)
    if gap_ids != EXPECTED_PRODUCT_GAP_IDS:
        raise AuditError(
            f"production semantic gap set drift: {gap_ids!r} != "
            f"{EXPECTED_PRODUCT_GAP_IDS!r}")

    comparison_result = helper.get("comparison")
    consolidation_result = helper.get("consolidation")
    if not isinstance(comparison_result, dict) or not isinstance(
            consolidation_result, dict):
        raise AuditError("product helper result shape drift")
    counts = comparison_result.get("counts")
    expected_counts = {
        "known": True,
        "paired_rows": 31,
        "side_a_only_rows": 1,
        "side_b_only_rows": 0,
        "differing_rows": 26,
        "differing_cells": 26,
    }
    if not isinstance(counts, dict) or any(
            counts.get(key) != value for key, value in expected_counts.items()):
        raise AuditError(f"product helper comparison counts drift: {counts!r}")
    if (comparison_result.get("status"), comparison_result.get("completion"),
            comparison_result.get("verdict"), comparison_result.get("skipped_inputs"),
            comparison_result.get("failed_inputs")) != (
                "ok", "complete", "diff", 0, 0):
        raise AuditError("product helper comparison outcome drift")
    if comparison_result.get("warnings") or comparison_result.get("failures"):
        raise AuditError("product helper unexpectedly reported warnings/failures")
    if (consolidation_result.get("status"), consolidation_result.get("completion"),
            consolidation_result.get("skipped_inputs"),
            consolidation_result.get("failed_inputs")) != (
                "ok", "complete", 0, 0):
        raise AuditError("product helper consolidation outcome drift")

    return {
        "formulas": {
            "package_without_core": _zip_digest_without_core(formulas_path),
            "semantic_digest": _workbook_semantic_digest(formulas_path),
        },
        "values": {
            "package_without_core": _zip_digest_without_core(values_path),
            "semantic_digest": _workbook_semantic_digest(values_path),
        },
        "product_rows": product_rows,
        "counts": expected_counts,
        "semantic_gaps": gaps,
        "semantic_gap_ids": list(gap_ids),
        "familiar_sheet": {
            "numeric_rows_exact": True,
            "source_backed_rows": 29,
            "placeholder_zero_rows": [P_KEY, V_KEY],
            "display_only_footnote_exact": True,
            "note_claims": note_claims,
            "note_contradicts_comparison_rows": True,
        },
        "provenance": {
            "summary_source_note": source_note,
            "familiar_source_note": familiar_source_note,
            "saved_notes_use_basenames_only": (
                "ramp_summary_consolidated.xlsx" in source_note
                and "tsn_ramp_summary_normalized.xlsx" in source_note
                and "\\" not in source_note and "/" not in source_note),
            "printed_source_provenance_presence": printed_presence,
            "printed_source_provenance_preserved": any(printed_presence.values()),
        },
        "source_backed_values_exact": True,
        "comparison_semantics_exact": False,
    }


def _sanitized_helper_result(payload: dict[str, object]) -> dict[str, object]:
    comparison = payload["comparison"]
    consolidation = payload["consolidation"]
    generation = comparison.get("artifact_generation") or {}
    members = sorted(
        ({
            "flavor": member.get("flavor"),
            "commit_role": member.get("commit_role"),
            "filename": Path(str(member.get("path"))).name,
        } for member in generation.get("members", [])),
        key=lambda item: str(item["flavor"]),
    )
    return {
        "consolidation": {
            key: consolidation.get(key)
            for key in ("status", "completion", "skipped_inputs", "failed_inputs")
        },
        "comparison": {
            key: comparison.get(key)
            for key in ("status", "completion", "verdict", "skipped_inputs",
                        "failed_inputs", "counts", "warnings", "failures")
        },
        "artifact_generation": {
            "completion": generation.get("completion"),
            "publication_state": generation.get("publication_state"),
            "requested_mode": generation.get("requested_mode"),
            "members": members,
        },
    }


def _parse_helper_stdout(stdout: str) -> dict[str, object]:
    for line in reversed([line.strip() for line in stdout.splitlines() if line.strip()]):
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict) and "outputs" in payload:
            return payload
    raise AuditError("product helper emitted no structured JSON result")


def _run_product_once(
        run_root: Path, pdf_root: Path, tsn_xlsx: Path,
        raw_records: dict[str, dict[str, int]], truth: dict[str, object],
        tsmis_aggregate: dict[str, int], tsn_counts: dict[str, int]
) -> dict[str, object]:
    environment = dict(os.environ)
    environment["PYTHONIOENCODING"] = "utf-8"
    completed = subprocess.run(
        [
            sys.executable, str(PRODUCT_HELPER_PATH),
            "--pdf-root", str(pdf_root),
            "--tsn-xlsx", str(tsn_xlsx),
            "--work-root", str(run_root),
        ],
        cwd=REPO_ROOT,
        env=environment,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=300,
        check=False,
    )
    if completed.returncode != 0:
        raise AuditError(
            f"product helper failed ({completed.returncode}): "
            f"{completed.stderr[-2000:]!r} {completed.stdout[-2000:]!r}")
    payload = _parse_helper_stdout(completed.stdout)
    outputs = payload.get("outputs")
    if not isinstance(outputs, dict):
        raise AuditError("product helper output-path map missing")
    expected_paths = {
        "consolidated": run_root / "ramp_summary_consolidated.xlsx",
        "formulas": run_root / "ramp_summary_comparison.xlsx",
        "values": run_root / "ramp_summary_comparison (values).xlsx",
    }
    for label, expected in expected_paths.items():
        observed = Path(str(outputs.get(label))).resolve()
        if observed != expected.resolve() or not observed.is_file():
            raise AuditError(
                f"product helper {label} path escaped/drifted: {observed}")
    consolidated = _inspect_production_consolidated(
        expected_paths["consolidated"], raw_records)
    comparison = _inspect_production_comparison(
        expected_paths["formulas"], expected_paths["values"], truth,
        tsmis_aggregate, tsn_counts, payload)
    product_code = payload.get("loaded_product_code")
    if not isinstance(product_code, dict) or not product_code.get("entries"):
        raise AuditError("product helper loaded-code manifest missing")
    return {
        "helper_result": _sanitized_helper_result(payload),
        "consolidated": consolidated,
        "comparison": comparison,
        "loaded_product_code": product_code,
    }


def _run_product(
        work_root: Path, pdf_root: Path, tsn_xlsx: Path,
        raw_records: dict[str, dict[str, int]], truth: dict[str, object],
        tsmis_aggregate: dict[str, int], tsn_counts: dict[str, int]
) -> dict[str, object]:
    first = _run_product_once(
        work_root / "product_a", pdf_root, tsn_xlsx, raw_records, truth,
        tsmis_aggregate, tsn_counts)
    second = _run_product_once(
        work_root / "product_b", pdf_root, tsn_xlsx, raw_records, truth,
        tsmis_aggregate, tsn_counts)
    if first != second:
        first_digest = _sha_bytes(json.dumps(
            first, sort_keys=True, ensure_ascii=False,
            separators=(",", ":")).encode("utf-8"))
        second_digest = _sha_bytes(json.dumps(
            second, sort_keys=True, ensure_ascii=False,
            separators=(",", ":")).encode("utf-8"))
        raise AuditError(
            f"production stable replay differs: {first_digest} != {second_digest}")
    return {
        "replays": 2,
        "stable_replay_exact": True,
        "stable_replay_sha256": _sha_bytes(json.dumps(
            first, sort_keys=True, ensure_ascii=False,
            separators=(",", ":")).encode("utf-8")),
        **first,
    }


def _detail_reconciliation(
        summary_records: dict[str, dict[str, int]],
        detail_xlsx_counts: dict[str, int],
        detail_pdf_counts: dict[str, int]) -> dict[str, object]:
    rows = []
    mismatches = []
    for route, record in summary_records.items():
        row = {
            "route": route,
            "summary_total": record["total_ramps"],
            "detail_xlsx_rows": detail_xlsx_counts.get(route),
            "detail_pdf_rows": detail_pdf_counts.get(route),
        }
        row["all_equal"] = (
            row["summary_total"] == row["detail_xlsx_rows"]
            == row["detail_pdf_rows"])
        rows.append(row)
        if not row["all_equal"]:
            mismatches.append(row)
    extra_xlsx = sorted(set(detail_xlsx_counts) - set(summary_records))
    extra_pdf = sorted(set(detail_pdf_counts) - set(summary_records))
    return {
        "routes": rows,
        "mismatches": mismatches,
        "extra_detail_xlsx_routes": extra_xlsx,
        "extra_detail_pdf_routes": extra_pdf,
        "all_exact": not mismatches and not extra_xlsx and not extra_pdf,
        "ordered_typed_sha256": _canonical_digest(
            (row["route"], row["summary_total"], row["detail_xlsx_rows"],
             row["detail_pdf_rows"])
            for row in rows),
    }


def _file_identity(path: Path) -> dict[str, object]:
    return {"bytes": path.stat().st_size, "sha256": _sha_file(path)}


def _loaded_oracle_module_manifest() -> dict[str, object]:
    entries = []
    for module_name, module in sorted(sys.modules.items()):
        if module_name.split(".", 1)[0] not in {
                "pdfplumber", "pdfminer", "pypdf", "openpyxl"}:
            continue
        raw = getattr(module, "__file__", None)
        if not raw:
            continue
        path = Path(raw).resolve()
        if path.suffix.lower() != ".py" or not path.is_file():
            continue
        entries.append({
            "module": module_name,
            "filename": path.name,
            "bytes": path.stat().st_size,
            "sha256": _sha_file(path),
        })
    canonical = json.dumps(
        entries, sort_keys=True, ensure_ascii=False,
        separators=(",", ":")).encode("utf-8")
    return {
        "module_file_count": len(entries),
        "canonical_json_sha256": _sha_bytes(canonical),
        "entries": entries,
    }


def _product_manifest_current(manifest: dict[str, object]) -> dict[str, object]:
    checks = []
    entries = manifest.get("entries")
    if not isinstance(entries, list):
        return {"all_current": False, "checks": [], "reason": "entries missing"}
    for entry in entries:
        relative = str(entry.get("relative_path", ""))
        path = (REPO_ROOT / "scripts" / relative).resolve()
        expected = {"bytes": entry.get("bytes"), "sha256": entry.get("sha256")}
        observed = _file_identity(path) if path.is_file() else None
        checks.append({
            "module": entry.get("module"),
            "relative_path": relative,
            "expected": expected,
            "observed": observed,
            "current": observed == expected,
        })
    return {
        "all_current": bool(checks) and all(check["current"] for check in checks),
        "checks": checks,
    }


def _run_gate(path: Path) -> dict[str, object]:
    completed = subprocess.run(
        [sys.executable, str(path)],
        cwd=REPO_ROOT,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=180,
        check=False,
        env={**os.environ, "PYTHONIOENCODING": "utf-8"},
    )
    if completed.returncode != 0:
        raise AuditError(
            f"mutation gate failed ({completed.returncode}): "
            f"{completed.stdout[-2000:]!r} {completed.stderr[-2000:]!r}")
    return {
        "status": "executed_pass",
        "gate_identity": _file_identity(path),
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
    }


def _mutation_probes(truth: dict[str, object]) -> dict[str, object]:
    probes = []

    def record(name: str, detected: bool) -> None:
        probes.append({"name": name, "detected": bool(detected)})

    for label, value in (("boolean", True), ("float", 1.0),
                         ("string", "1"), ("negative", -1)):
        try:
            _strict_count(value, f"probe {label}")
        except AuditError:
            record(f"strict_count_rejects_{label}", True)
        else:
            record(f"strict_count_rejects_{label}", False)

    exact = {"001": {"count": 1}}
    changed = {"001": {"count": 2}}
    record("cross_format_count_mutation",
           _cross_format(exact, changed)["difference_count"] == 1)
    record("suffix_route_identity_preserved",
           _route_from_name(Path("tsar_ramp_summary_route_005S.pdf"), ".pdf")
           == "005S")

    synthetic = {slug: 0 for slug, _row, _label in ALL_SOURCE}
    synthetic.update({"total_ramps": 1, "ramp_points_no_linework": 0})
    arithmetic = _arithmetic({"001": synthetic}, {})
    record("unexplained_ramp_type_residual",
           len(arithmetic["unexplained_after_detail_pv"]) == 1)

    product_rows = []
    for row in truth["rows"]:
        product_rows.append({
            "category": row["category"],
            "status": "Both",
            "tsmis": 0 if row["category"] in (P_KEY, V_KEY) else row["tsmis"],
            "tsn": row["tsn"],
        })
    product_rows.append({
        "category": "Ramp Points w/out linework", "status": "TSMIS only",
        "tsmis": 59, "tsn": None,
    })
    record("exact_three_product_semantic_gaps",
           tuple(gap["id"] for gap in _semantic_gaps(
               product_rows, truth["rows"])) == EXPECTED_PRODUCT_GAP_IDS)
    return {
        "probes": probes,
        "all_detected": all(probe["detected"] for probe in probes),
    }


def _publication_current(args: argparse.Namespace,
                         result: dict[str, object]) -> tuple[bool, dict[str, object]]:
    tree_paths = {
        "summary_pdf": args.summary_pdf_root,
        "summary_xlsx": args.summary_xlsx_root,
        "detail_xlsx": args.detail_xlsx_root,
        "detail_pdf": args.detail_pdf_root,
    }
    file_paths = {
        "tsn_raw_pdf": args.tsn_raw,
        "tsn_normalized_xlsx": args.tsn_xlsx,
        "stage6_result": args.stage6_result,
        "stage6_acceptance": args.stage6_acceptance,
    }
    detail: dict[str, object] = {"trees": {}, "files": {}, "code": {}}
    captures = result["source_capture"]
    for label, path in tree_paths.items():
        observed, _entries = _manifest(path, str(TREE_BINDINGS[label]["suffix"]))
        expected = captures[label]["observed"]
        detail["trees"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    for label, path in file_paths.items():
        observed = _file_identity(path) if path.is_file() else None
        expected = captures[label]["observed"]
        detail["files"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    code_paths = {
        "generator": GENERATOR_PATH,
        "product_helper": PRODUCT_HELPER_PATH,
        "self_gate": SELF_GATE_PATH,
    }
    for label, path in code_paths.items():
        observed = _file_identity(path) if path.is_file() else None
        expected = result["provenance"]["code_identities"][label]
        detail["code"][label] = {
            "expected": expected, "observed": observed,
            "current": observed == expected,
        }
    parser_manifest = _loaded_oracle_module_manifest()
    expected_parser = result["provenance"]["loaded_oracle_module_manifest"]
    detail["code"]["oracle_parser_modules"] = {
        "expected_sha256": expected_parser["canonical_json_sha256"],
        "observed_sha256": parser_manifest["canonical_json_sha256"],
        "expected_files": expected_parser["module_file_count"],
        "observed_files": parser_manifest["module_file_count"],
        "current": parser_manifest == expected_parser,
    }
    product_current = _product_manifest_current(
        result["production"]["loaded_product_code"])
    detail["code"]["loaded_product_modules"] = product_current
    current_flags = [
        item["current"]
        for group in (detail["trees"], detail["files"], detail["code"])
        for item in group.values()
        if "current" in item
    ]
    current_flags.append(product_current["all_current"])
    return all(current_flags), detail


def run(args: argparse.Namespace) -> dict[str, object]:
    code_identities = {
        "generator": _file_identity(GENERATOR_PATH),
        "product_helper": _file_identity(PRODUCT_HELPER_PATH),
        "self_gate": _file_identity(SELF_GATE_PATH),
    }
    mutation_gate = _run_gate(SELF_GATE_PATH)

    work_root = args.work_root.resolve()
    source_roots = [
        args.summary_pdf_root.resolve(), args.summary_xlsx_root.resolve(),
        args.detail_xlsx_root.resolve(), args.detail_pdf_root.resolve(),
    ]
    if any(work_root == source or work_root.is_relative_to(source)
           for source in source_roots):
        raise AuditError("private work root must not be inside an authoritative source tree")
    work_root.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(
            prefix="ramp-summary-stage8-", dir=work_root) as temporary:
        snapshot_root = Path(temporary).resolve()
        if snapshot_root.parent != work_root:
            raise AuditError("private snapshot escaped the requested work root")

        source_capture: dict[str, dict[str, object]] = {}
        source_capture["summary_pdf"], summary_pdf_root = _capture_tree(
            "summary_pdf", args.summary_pdf_root, TREE_BINDINGS["summary_pdf"],
            snapshot_root / "summary_pdf")
        source_capture["summary_xlsx"], summary_xlsx_root = _capture_tree(
            "summary_xlsx", args.summary_xlsx_root, TREE_BINDINGS["summary_xlsx"],
            snapshot_root / "summary_xlsx")
        source_capture["detail_xlsx"], detail_xlsx_root = _capture_tree(
            "detail_xlsx", args.detail_xlsx_root, TREE_BINDINGS["detail_xlsx"],
            snapshot_root / "detail_xlsx")
        source_capture["detail_pdf"], detail_pdf_root = _capture_tree(
            "detail_pdf", args.detail_pdf_root, TREE_BINDINGS["detail_pdf"],
            snapshot_root / "detail_pdf")
        source_capture["tsn_raw_pdf"], tsn_raw = _capture_file(
            "tsn_raw_pdf", args.tsn_raw, FILE_BINDINGS["tsn_raw_pdf"],
            snapshot_root / "tsn_raw.pdf")
        source_capture["tsn_normalized_xlsx"], tsn_xlsx = _capture_file(
            "tsn_normalized_xlsx", args.tsn_xlsx,
            FILE_BINDINGS["tsn_normalized_xlsx"],
            snapshot_root / "tsn_normalized.xlsx")
        source_capture["stage6_result"], stage6_result = _capture_file(
            "stage6_result", args.stage6_result, FILE_BINDINGS["stage6_result"],
            snapshot_root / "stage6_result.json")
        source_capture["stage6_acceptance"], stage6_acceptance = _capture_file(
            "stage6_acceptance", args.stage6_acceptance,
            FILE_BINDINGS["stage6_acceptance"],
            snapshot_root / "stage6_acceptance.json")

        summary_pdf = _parse_summary_pdfs(summary_pdf_root)
        summary_xlsx = _parse_summary_xlsx(summary_xlsx_root)
        detail_xlsx = _parse_detail_xlsx(detail_xlsx_root)
        detail_pdf = _parse_detail_pdfs(detail_pdf_root)
        route_universes = _require_exact_route_universes({
            "summary_pdf": summary_pdf["routes"],
            "summary_xlsx": summary_xlsx["routes"],
            "detail_xlsx": detail_xlsx["routes"],
            "detail_pdf": detail_pdf["routes"],
        })
        cross_format = _cross_format(
            summary_pdf["records"], summary_xlsx["records"])
        arithmetic = _arithmetic(
            summary_pdf["records"], detail_pdf["pv_by_route"])
        aggregate = _aggregate(summary_pdf["records"])
        detail_reconciliation = _detail_reconciliation(
            summary_pdf["records"], detail_xlsx["route_counts"],
            detail_pdf["route_counts"])
        tsn = _load_tsn_normalized(tsn_xlsx)
        stage6 = _load_stage6(stage6_result, stage6_acceptance)
        truth = _comparison_truth(aggregate, tsn["counts"])
        mutations = _mutation_probes(truth)
        production = _run_product(
            snapshot_root, summary_pdf_root, tsn_xlsx, summary_pdf["records"],
            truth, aggregate, tsn["counts"])

    parser_manifest = _loaded_oracle_module_manifest()
    product_code_current = _product_manifest_current(
        production["loaded_product_code"])
    expected_residual_routes = list(EXPECTED_PV_BY_ROUTE)
    residual_routes = [
        row["route"] for row in arithmetic["unprinted_ramp_type_routes"]]
    source_invariants = {
        "all_four_source_trees_match_exact_bound_manifests": all(
            source_capture[label]["observed"] == {
                key: TREE_BINDINGS[label][key]
                for key in ("files", "bytes", "manifest_sha256")
            } | {"serialization": "name\\tbytes\\tsha256\\n sorted by name"}
            for label in TREE_BINDINGS),
        "all_four_route_universes_exact_ordered_unique": route_universes["all_exact"],
        "summary_pdf_producer_and_generation_metadata_exact": (
            summary_pdf["metadata_signature_count"] == 95
            and sum(item["count"] for item in summary_pdf["metadata_signatures"]) == 126),
        "summary_pdf_xlsx_all_3780_values_exact": (
            cross_format["compared_values"] == 3_780
            and cross_format["difference_count"] == 0
            and cross_format["ordered_typed_sha256"]
            == EXPECTED_CROSS_FORMAT_SHA256),
        "summary_aggregate_exact": aggregate == EXPECTED_TSMIS_AGGREGATE,
        "summary_detail_xlsx_pdf_route_counts_exact": (
            detail_reconciliation["all_exact"]
            and detail_xlsx["rows"] == 15_216
            and detail_pdf["rows"] == 15_216),
        "detail_pdf_page_universe_exact": (
            detail_pdf["pages"] == 626 and detail_pdf["data_pages"] == 500),
        "detail_pv_records_exact": (
            detail_pdf["pv_by_route"] == EXPECTED_PV_BY_ROUTE
            and detail_pdf["pv_by_type"] == EXPECTED_PV_BY_TYPE
            and len(detail_pdf["pv_records"]) == 22),
        "all_summary_axes_conserve": (
            not arithmetic["nonzero_highway"]
            and not arithmetic["nonzero_onoff"]
            and not arithmetic["nonzero_population"]),
        "unprinted_ramp_type_residual_exactly_explained_by_detail_pv": (
            residual_routes == expected_residual_routes
            and not arithmetic["unexplained_after_detail_pv"]),
        "tsn_raw_normalized_stage6_chain_exact": all(stage6["checks"].values()),
        "comparison_truth_shape_and_digest_exact": (
            truth["both"] == 29 and truth["only_in_tsn"] == 2
            and truth["only_in_tsmis"] == 0
            and truth["identical_both"] == 5
            and truth["differing_both"] == 24
            and truth["ordered_typed_sha256"]
            == EXPECTED_COMPARISON_TRUTH_SHA256),
        "comparison_totals_exact": (
            aggregate["total_ramps"] == 15_216
            and tsn["counts"]["Total Number of Ramps"] == 15_410
            and tsn["counts"]["Total Number of Ramps"]
            - aggregate["total_ramps"] == 194),
    }
    projection_exact = bool(
        production["consolidated"]["projection_exact"]
        and production["comparison"]["source_backed_values_exact"])
    gap_ids = tuple(production["comparison"]["semantic_gap_ids"])
    audit_invariants = {
        **source_invariants,
        "semantic_mutations_all_detected": mutations["all_detected"],
        "permanent_mutation_gate_executed_pass": (
            mutation_gate["status"] == "executed_pass"),
        "production_two_replays_semantically_and_package_stable": (
            production["replays"] == 2 and production["stable_replay_exact"]),
        "production_source_backed_value_projection_exact": projection_exact,
        "production_gap_set_exactly_three_and_documented": (
            gap_ids == EXPECTED_PRODUCT_GAP_IDS),
        "loaded_product_code_hash_manifest_current_at_result_build": (
            product_code_current["all_current"]),
        "loaded_oracle_parser_manifest_nonempty": (
            parser_manifest["module_file_count"] > 0),
    }

    result = {
        "schema_version": 1,
        "audit": "Stage 8 Ramp Summary authoritative TSMIS-vs-TSN base comparison oracle",
        "methodology": {
            "authority": (
                "Exact All Reports 7.9 same-pull TSMIS Ramp Summary PDF/XLSX bytes, "
                "same-pull Ramp Detail PDF/XLSX rows, and the exact accepted TSN "
                "raw-to-r7-normalized chain."),
            "independence": (
                "Truth-side parsing imports no application parser, normalizer, "
                "schema, comparator, or writer. Production imports occur only in "
                "an isolated child process after truth is derived."),
            "outcomes_separated": [
                "source_truth_exact", "production_value_projection_exact",
                "production_comparison_semantics_exact",
                "stage8_base_oracle_complete", "comparison_end_to_end_perfect",
            ],
            "comparison_contract": (
                "29 shared categories compare; P and V are TSN-only; the 59-point "
                "no-linework metric is display-only and excluded from verdict inputs."),
            "visual_verification": VISUAL_REVIEW,
        },
        "bindings": {
            "source_trees": TREE_BINDINGS,
            "files": FILE_BINDINGS,
            "expected_cross_format_sha256": EXPECTED_CROSS_FORMAT_SHA256,
            "expected_comparison_truth_sha256": EXPECTED_COMPARISON_TRUTH_SHA256,
        },
        "source_capture": source_capture,
        "route_universes": route_universes,
        "tsmis_summary": {
            "pdf": {
                "routes": summary_pdf["routes"],
                "pages": summary_pdf["pages"],
                "metadata_signatures": summary_pdf["metadata_signatures"],
                "metadata_signature_count": summary_pdf["metadata_signature_count"],
                "metadata_ordered_typed_sha256": summary_pdf[
                    "metadata_ordered_typed_sha256"],
                "geometry_typed_sha256": summary_pdf["geometry_typed_sha256"],
                "records": summary_pdf["records"],
            },
            "xlsx": {
                "routes": summary_xlsx["routes"],
                "generated_timestamps": summary_xlsx["generated_timestamps"],
            },
            "cross_format": cross_format,
            "aggregate": aggregate,
            "route_arithmetic": arithmetic,
        },
        "same_pull_ramp_detail": {
            "xlsx": {
                "routes": detail_xlsx["routes"],
                "route_counts": detail_xlsx["route_counts"],
                "rows": detail_xlsx["rows"],
                "ordered_typed_sha256": detail_xlsx["ordered_typed_sha256"],
            },
            "pdf": detail_pdf,
            "summary_detail_reconciliation": detail_reconciliation,
        },
        "tsn": {
            "normalized": tsn,
            "stage6_chain": stage6,
        },
        "intended_comparison_truth": truth,
        "production": production,
        "semantic_mutation_probes": mutations,
        "dependency_gates": {
            "stage8_ramp_summary_mutations": mutation_gate,
        },
        "provenance": {
            "code_identities": code_identities,
            "loaded_oracle_module_manifest": parser_manifest,
            "loaded_product_code_current_at_result_build": product_code_current,
        },
        "findings": {
            "oracle_blocking": [],
            "product_red": [
                {
                    "finding": "CMP-AUD-019",
                    "fact": (
                        "Production reports consolidation complete with zero warnings "
                        "although nine authentic route audit formulas evaluate to the "
                        "Ramp Types source-residual warning."),
                    "evidence": production["consolidated"]["audit_warning_routes"],
                },
                {
                    "finding": "CMP-AUD-020",
                    "fact": (
                        "The current source happens to reconcile exactly, but the product "
                        "does not make all route-axis/detail conservation a completion gate."),
                    "evidence": {
                        "routes": 126, "detail_rows": 15_216,
                        "pv_residual_records": 22,
                    },
                },
                {
                    "finding": "CMP-AUD-024",
                    "fact": (
                        "The 59-point no-linework display metric is injected as an Only in "
                        "TSMIS comparison row and therefore participates in the verdict."),
                    "evidence": production["comparison"]["semantic_gaps"][2],
                },
                {
                    "finding": "CMP-AUD-025",
                    "fact": (
                        "P and V are source-proven TSN-only categories, but production "
                        "fabricates TSMIS zeros and labels both rows shared."),
                    "evidence": production["comparison"]["semantic_gaps"][:2],
                },
                {
                    "finding": "CMP-AUD-071",
                    "fact": (
                        "The authentic 126-route production output is exact, but current "
                        "product completion does not enforce this exact route universe."),
                    "evidence": route_universes["universes"],
                },
                {
                    "finding": "CMP-AUD-076",
                    "fact": (
                        "Production retains source basenames but loses TSMIS report date, "
                        "reference date, submitter, and generated-time provenance."),
                    "evidence": production["comparison"]["provenance"],
                },
                {
                    "finding": "CMP-AUD-146",
                    "fact": (
                        "The accepted TSN normalized Summary preserves counts but not all "
                        "printed TSN report identity/timing provenance."),
                    "evidence": stage6,
                },
            ],
        },
        "audit_invariants": audit_invariants,
        "source_truth_exact": all(source_invariants.values()),
        "production_value_projection_exact": projection_exact,
        "production_comparison_semantics_exact": False,
        "stage8_base_oracle_complete": all(audit_invariants.values()),
        "comparison_end_to_end_perfect": False,
    }
    current, current_detail = _publication_current(args, result)
    result["provenance"]["final_revalidation_at_result_build"] = {
        "all_current": current,
        "detail": current_detail,
    }
    result["audit_invariants"]["sources_and_code_current_at_result_build"] = current
    result["source_truth_exact"] = bool(result["source_truth_exact"] and current)
    result["stage8_base_oracle_complete"] = all(
        result["audit_invariants"].values())
    return result


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
                mode="w", encoding="utf-8", newline="\n", delete=False,
                dir=path.parent, prefix=f".{path.name}.", suffix=".tmp") as handle:
            temporary = Path(handle.name)
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary is not None:
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass


def _unlink_if_present(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass


def _write_decision(path: Path, output: Path, result: dict[str, object],
                    *, accepted: bool, reason: str,
                    postwrite_current: bool,
                    postwrite_detail: dict[str, object],
                    open_findings_authorized: bool) -> dict[str, object]:
    result_identity = _file_identity(output)
    decision = {
        "schema_version": 1,
        "accepted": accepted,
        "reason": reason,
        "audit": result.get("audit"),
        "result": str(output.resolve()),
        "result_bytes": result_identity["bytes"],
        "result_sha256": result_identity["sha256"],
        "source_truth_exact": result.get("source_truth_exact", False),
        "production_value_projection_exact": result.get(
            "production_value_projection_exact", False),
        "production_comparison_semantics_exact": result.get(
            "production_comparison_semantics_exact", False),
        "stage8_base_oracle_complete": result.get(
            "stage8_base_oracle_complete", False),
        "comparison_end_to_end_perfect": result.get(
            "comparison_end_to_end_perfect", False),
        "open_product_findings_authorized": open_findings_authorized,
        "post_result_write_revalidation": postwrite_current,
        "post_result_write_identities": postwrite_detail,
    }
    _atomic_write_text(path, json.dumps(
        decision, indent=2, ensure_ascii=False) + "\n")
    return decision


def _parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--summary-pdf-root", type=Path, default=DEFAULT_SUMMARY_PDF_ROOT)
    parser.add_argument(
        "--summary-xlsx-root", type=Path, default=DEFAULT_SUMMARY_XLSX_ROOT)
    parser.add_argument(
        "--detail-xlsx-root", type=Path, default=DEFAULT_DETAIL_XLSX_ROOT)
    parser.add_argument(
        "--detail-pdf-root", type=Path, default=DEFAULT_DETAIL_PDF_ROOT)
    parser.add_argument("--tsn-raw", type=Path, default=DEFAULT_TSN_RAW)
    parser.add_argument("--tsn-xlsx", type=Path, default=DEFAULT_TSN_XLSX)
    parser.add_argument(
        "--stage6-result", type=Path, default=DEFAULT_STAGE6_RESULT)
    parser.add_argument(
        "--stage6-acceptance", type=Path, default=DEFAULT_STAGE6_ACCEPTANCE)
    parser.add_argument("--work-root", type=Path, default=DEFAULT_WORK_ROOT)
    parser.add_argument(
        "--allow-open-findings", action="store_true",
        help=(
            "accept the completed oracle/value projection while the exact documented "
            "product comparison-semantic findings remain open"))
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(sys.argv[1:] if argv is None else argv)
    acceptance_path = args.output.with_suffix(args.output.suffix + ".acceptance.json")
    rejection_path = args.output.with_suffix(args.output.suffix + ".rejection.json")
    _unlink_if_present(acceptance_path)
    _unlink_if_present(rejection_path)
    try:
        result = run(args)
    except Exception as exc:
        failure = {
            "schema_version": 1,
            "audit": "Stage 8 Ramp Summary authoritative TSMIS-vs-TSN base comparison oracle",
            "error": {"type": type(exc).__name__, "message": str(exc)},
            "source_truth_exact": False,
            "production_value_projection_exact": False,
            "production_comparison_semantics_exact": False,
            "stage8_base_oracle_complete": False,
            "comparison_end_to_end_perfect": False,
        }
        _atomic_write_text(args.output, json.dumps(
            failure, indent=2, ensure_ascii=False) + "\n")
        decision = _write_decision(
            rejection_path, args.output, failure, accepted=False,
            reason="oracle_execution_failed", postwrite_current=False,
            postwrite_detail={}, open_findings_authorized=False)
        sys.stdout.write(json.dumps({
            "accepted": False,
            "reason": decision["reason"],
            "output": str(args.output),
            "rejection": str(rejection_path),
            "error": failure["error"],
        }, ensure_ascii=False) + "\n")
        return 2

    prewrite_current, prewrite_detail = _publication_current(args, result)
    result["publication_revalidation"] = {
        "after_complete_result_build": True,
        "before_result_write_all_current": prewrite_current,
        "before_result_write_identities": prewrite_detail,
    }
    result["audit_invariants"]["publication_inputs_current_before_write"] = (
        prewrite_current)
    result["stage8_base_oracle_complete"] = all(
        result["audit_invariants"].values())
    _atomic_write_text(args.output, json.dumps(
        result, indent=2, ensure_ascii=False) + "\n")

    postwrite_current, postwrite_detail = _publication_current(args, result)
    if not postwrite_current:
        result["publication_revalidation"]["post_result_write_all_current"] = False
        result["publication_revalidation"]["post_result_write_identities"] = (
            postwrite_detail)
        result["audit_invariants"]["publication_inputs_current_after_write"] = False
        result["stage8_base_oracle_complete"] = False
        _atomic_write_text(args.output, json.dumps(
            result, indent=2, ensure_ascii=False) + "\n")
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False,
            reason="post_result_write_revalidation_failed",
            postwrite_current=False, postwrite_detail=postwrite_detail,
            open_findings_authorized=False)
        sys.stdout.write(json.dumps({
            "accepted": False, "reason": decision["reason"],
            "rejection": str(rejection_path),
        }, ensure_ascii=False) + "\n")
        return 2

    open_findings = not result["production_comparison_semantics_exact"]
    accepted = bool(
        result["source_truth_exact"]
        and result["production_value_projection_exact"]
        and result["stage8_base_oracle_complete"]
        and postwrite_current
        and (not open_findings or args.allow_open_findings))
    if not accepted:
        reason = (
            "open_product_findings_not_authorized"
            if (result["stage8_base_oracle_complete"] and open_findings
                and not args.allow_open_findings)
            else "audit_or_projection_incomplete"
        )
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False, reason=reason,
            postwrite_current=postwrite_current,
            postwrite_detail=postwrite_detail,
            open_findings_authorized=False)
        sys.stdout.write(json.dumps({
            "accepted": False,
            "reason": reason,
            "output": str(args.output),
            "result_bytes": decision["result_bytes"],
            "result_sha256": decision["result_sha256"],
            "rejection": str(rejection_path),
        }, ensure_ascii=False) + "\n")
        return 1 if reason == "open_product_findings_not_authorized" else 2

    decision = _write_decision(
        acceptance_path, args.output, result, accepted=True,
        reason="oracle_complete_with_documented_product_findings",
        postwrite_current=postwrite_current, postwrite_detail=postwrite_detail,
        open_findings_authorized=bool(args.allow_open_findings and open_findings))
    acceptance_identity = _file_identity(acceptance_path)
    sys.stdout.write(json.dumps({
        "accepted": True,
        "output": str(args.output),
        "result_bytes": decision["result_bytes"],
        "result_sha256": decision["result_sha256"],
        "acceptance": str(acceptance_path),
        "acceptance_bytes": acceptance_identity["bytes"],
        "acceptance_sha256": acceptance_identity["sha256"],
        "source_truth_exact": result["source_truth_exact"],
        "production_value_projection_exact": result[
            "production_value_projection_exact"],
        "production_comparison_semantics_exact": result[
            "production_comparison_semantics_exact"],
        "stage8_base_oracle_complete": result["stage8_base_oracle_complete"],
        "product_findings": len(result["findings"]["product_red"]),
    }, ensure_ascii=False) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
