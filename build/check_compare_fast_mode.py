"""Regression lock for the opt-in Fast vs-TSN workbook serializer.

The fast path may reduce openpyxl descriptor work, but it is accepted only when
both the typed comparison truth and every stable OOXML package member are exact
matches for the historical writer.  docProps/core.xml is excluded because its
created/modified timestamps are intentionally run-specific.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT / "build"), str(ROOT)]

import compare_core as core  # noqa: E402
from benchmark_vs_tsn_speed import _package_manifest  # noqa: E402


_failures = []


def check(name, condition, detail=""):
    print(f"  [{'OK ' if condition else 'FAIL'}] {name}")
    if not condition:
        _failures.append(name)
        if detail:
            print(f"       {detail}")


def _ditto_resolver(_rows, _has_route):
    # Exercises the post-construction fill mutation that requires each fast
    # cell to own its StyleArray instead of aliasing the cached prototype.
    return {0: {2: "paired value"}}


SCHEMA = core.CompareSchema(
    report_name="Fast Serializer Lock",
    header=["Loc", "Text", "Ditto", "Med Wid", "Context"],
    side_a="TSMIS",
    side_b="TSN",
    id_noun="row",
    id_noun_plural="rows",
    medwid_fields=("Med Wid",),
    context_fields=("Context",),
    ditto_nonasserting=True,
    ditto_resolver=_ditto_resolver,
)

LEFT = [
    ["A", "=1+1", "++", "8V", "left context"],
    ["B", "@literal", "same", "02", ""],
    ["ONLY-A", "#N/A", "+", "0", "context"],
]
RIGHT = [
    ["A", "=1+1", "++", "08V", "right context"],
    ["B", "different", "same", "2", ""],
    ["ONLY-B", "-literal", "+", "0", "context"],
]


def _binding_probe(fast_mode):
    wb = Workbook(write_only=True)
    wb._tsmis_fast_styles = fast_mode
    ws = wb.create_sheet("Probe")
    calls = []
    original = Cell._bind_value

    def counted(cell, value):
        calls.append(value)
        return original(cell, value)

    Cell._bind_value = counted
    try:
        cell = core._styled_cell(
            ws, "=1+1", font=Font(name="Arial", size=10), guard=True)
    finally:
        Cell._bind_value = original
        wb.close()
    return calls, cell.data_type


def _hot_loop_contract():
    wb = Workbook(write_only=True)
    wb._tsmis_fast_styles = True
    ws = wb.create_sheet("Identity")
    font = Font(name="Arial", size=10)
    same_value_distinct_font = Font(name="Arial", size=10)
    first = core._styled_cell(ws, "first", font=font)
    second = core._styled_cell(ws, "second", font=font)
    cache = wb._tsmis_fast_style_cache
    first_key = next(iter(cache))
    first_entry = cache[first_key]
    check("style cache keys are component identities",
          len(cache) == 1 and all(isinstance(item, int) for item in first_key)
          and first_entry[0] is font and first._style == second._style)
    third = core._styled_cell(ws, "third", font=same_value_distinct_font)
    check("equal-but-distinct components cannot alias an identity entry",
          font == same_value_distinct_font and len(cache) == 2
          and first._style == third._style)
    wb.close()

    standard_calls, standard_type = _binding_probe(False)
    fast_calls, fast_type = _binding_probe(True)
    check("Fast guarded literals bind once; Standard keeps its historical path",
          standard_calls == ["=1+1", "=1+1"]
          and fast_calls == ["=1+1"]
          and standard_type == fast_type == "s",
          f"standard={standard_calls!r}; fast={fast_calls!r}")


def _one_mode(root, mode):
    standard = root / f"{mode}-standard.xlsx"
    fast = root / f"{mode}-fast.xlsx"
    normal_result = core.run_compare(
        SCHEMA, LEFT, RIGHT, False, standard, mode=mode, fast_mode=False)
    fast_result = core.run_compare(
        SCHEMA, LEFT, RIGHT, False, fast, mode=mode, fast_mode=True)

    check(f"{mode}: both serializers build",
          normal_result.status == fast_result.status == "ok"
          and standard.is_file() and fast.is_file(),
          f"standard={normal_result.status!r}; fast={fast_result.status!r}")
    if not standard.is_file() or not fast.is_file():
        return

    normal_typed = normal_result.comparison_outcome.to_dict()
    fast_typed = fast_result.comparison_outcome.to_dict()
    check(f"{mode}: typed outcome is exact",
          normal_typed == fast_typed,
          f"standard={normal_typed!r}; fast={fast_typed!r}")

    normal_package = _package_manifest(standard)
    fast_package = _package_manifest(fast)
    check(f"{mode}: every stable OOXML member is byte-exact",
          normal_package == fast_package,
          f"standard={normal_package['sha256']}; fast={fast_package['sha256']}")
    check(f"{mode}: fixture covers guarded literals and per-cell style mutation",
          normal_typed["counts"]["differing_cells"] > 0
          and normal_package["member_count"] >= 15)


def main():
    print("Fast vs TSN serializer equivalence:")
    _hot_loop_contract()
    with tempfile.TemporaryDirectory(prefix="tsmis_fast_compare_") as tmp:
        root = Path(tmp)
        _one_mode(root, "values")
        _one_mode(root, "formulas")

    print()
    if _failures:
        print(f"FAILED: {len(_failures)} check(s): {_failures}")
        return 1
    print("ALL FAST-COMPARISON SERIALIZER CHECKS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
