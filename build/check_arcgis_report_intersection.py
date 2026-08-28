"""Golden check for the ArcGIS Intersection Detail lane (the second report
rendered from the layer library).

Highway Detail's lane is checked by `check_arcgis_report`; this one covers what
is DIFFERENT about a POINT report, and every rule here is one that was measured
against the real statewide export rather than assumed, so each has teeth:

  * the ROW UNIVERSE — an IM-managed shell (no inventory block) and a retired
    row are not records, and are counted as skipped rather than dropped
    silently;
  * the LEG RULE — the mainline block comes from the Major approach legs and the
    cross-street block from the Minor legs (measured 99.4-100% vs 62-96% for the
    reverse), and a disagreement among legs of one type prints the most frequent
    value;
  * the OVERLAY CARRY RULE — a gap in a span layer carries the previous value
    forward for Highway Group and Rural/Urban (continuous attributes) but NOT
    for City (containment: outside the limits there is no city). Measured
    94.5%/91.3% carry vs 98.1% strict for City;
  * the POSITION contract — the build writes the consolidated export's own
    column POSITIONS, whose header labels sit shifted against their values, so
    the shared by-position loaders read our build and the export identically;
  * the role gates — a swapped pair is refused rather than compared.

Driven from a synthetic layer library end to end through the shipped
`consolidate()`.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_arcgis_report_intersection.py
"""
import datetime
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import Workbook, load_workbook

import arcgis_report_intersection_detail as ari
import arcgis_reports
import clean_road_layers as crl
import compare_intersection_detail_arcgis as cmp_arc
import compare_intersection_detail_tsn as cid
import intersection_detail_columns as idc
import outcome

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _raises(fn, needle=None):
    try:
        fn()
    except Exception as e:
        return needle is None or needle.lower() in str(e).lower()
    return False


def _sheet(path, title, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(list(header))
    for r in rows:
        ws.append([r.get(c) for c in header])
    wb.save(path)


# --------------------------------------------------------------------------- #
# a synthetic layer library
# --------------------------------------------------------------------------- #
D = datetime.datetime


def span(begin, end, value, attr, county="ORA", route="001", prefix=None):
    return {"District": "District 12", "RouteNum": route, "RouteSuffix": ".",
            "Alignment": "Right", "BeginCounty": county, "EndCounty": county,
            "BeginPMPrefix": prefix or ".", "EndPMPrefix": prefix or ".",
            "BeginPMMeasure": begin, "EndPMMeasure": end,
            "LRSFromDate": D(2020, 1, 1), "LRSToDate": None,
            "RouteID": f"SHS_{route}._P", "BeginODMeasure": begin,
            "EndODMeasure": end, attr: value}


def inx(iid, pm, **over):
    r = {"INTERSECTION_ID": iid, "LRS_DATE_RETIRE": None,
         "InventoryItemStartDate": D(2023, 1, 1), "InventoryItemEndDate": None,
         "County_Code": "Orange", "District_Code": "District 12",
         "Main_RouteNum": "001", "Main_RouteSuffix": ".",
         "Main_PMPrefix": ".", "Main_PMSuffix": ".", "Main_PMMeasure": pm,
         "Main_Begin_Date": D(1973, 10, 19),
         "Cross_RouteNum": None, "Cross_RouteSuffix": ".",
         "Cross_PMPrefix": ".", "Cross_PMSuffix": ".", "Cross_PMMeasure": None,
         "Cross_Begin_Date": D(1974, 5, 1),
         "Intersection_Name": "MAIN ST   LT", "Intersection_Geometry": "T- Tee",
         "Int_Geometry_Begin_Date": D(1973, 10, 19),
         "Intersection_Control": "B- Stop Signs on Cross Street Only",
         "Int_Control_Begin_Date": D(1973, 10, 19),
         "Intersection_Lighted_Ind": "Yes",
         "Int_Lighted_Ind_Begin_Date": D(1973, 10, 19),
         "Int_Date_Of_Record": D(1973, 10, 19)}
    r.update(over)
    return r


def leg(approach_id, iid, leg_type, **over):
    seg = {"INTERSECTION_ID": iid, "APPROACH_ID": approach_id,
           "LEG_TYPE": leg_type}
    det = {"APPROACH_ID": approach_id, "Number_Thru_Lanes": 2,
           "FlowCode": "P- Two-Way Traffic", "N_Distance": 250,
           "Left_Channel": "N- No Left Turn Channelization",
           "Right_Channel_Ind": "No", "Signal_Arm_Ind": "No"}
    det.update(over)
    return seg, det


def build_library(root, *, inx_rows, legs, hg_spans, city_spans, pop_spans):
    """A minimal library: the five layers this report reads, plus the INDEX
    manifest the loader verifies them against (row counts included, so the
    truncated-export gate is exercised rather than bypassed)."""
    root.mkdir(parents=True, exist_ok=True)
    segs = [s for s, _d in legs]
    dets = [d for _s, d in legs]
    made = []
    for n, layer, cols, rows in (
            (38, ari.INX_LAYER, ari._INX_COLS, inx_rows),
            (37, ari.SEG_LAYER, ari._SEG_COLS, segs),
            (36, ari.APP_LAYER, ari._APP_COLS, dets)):
        name = f"{n}_{layer}.xlsx"
        _sheet(root / name, layer, cols, rows)
        made.append((name, layer, len(rows), len(cols)))
    for n, (tag, rows) in enumerate((("HG", hg_spans), ("CITY", city_spans),
                                     ("POP", pop_spans)), start=25):
        layer, attr, _carry = ari.OVERLAY_LAYERS[tag]
        cols = ari._SPAN_ID + [attr]
        name = f"{n}_{layer}.xlsx"
        _sheet(root / name, layer, cols, rows)
        made.append((name, layer, len(rows), len(cols)))
    _sheet(root / crl.INDEX_NAME, "INDEX", list(crl.INDEX_HEADER),
           [dict(zip(crl.INDEX_HEADER,
                     (fname, layer, nrows, nfields, str(root), "synthetic")))
            for fname, layer, nrows, nfields in made])


def read_out(path):
    wb = load_workbook(path, read_only=True)
    ws = wb[ari.SHEET_NAME]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) + [None] * (len(header) - len(r))
            for r in it if any(c is not None for c in r)]
    wb.close()
    return header, rows


def cell(row, name):
    return row[ari._POS[name]]


def main():
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_agrepinx_"))
    lib, out = tmp / "layers", tmp / "inx_from_layers.xlsx"
    asof = "2026-08-28"

    print("the projection contract:")
    check("every column position but Route is written",
          sorted(ari._POS.values()) == list(range(1, len(ari.HEADER))))
    check("PROJECTION and the position map agree",
          set(ari.PROJECTION) == set(ari._POS))
    check("the emitted header is exactly the consolidated export's",
          ari.HEADER == ["Route"] + list(idc.HEADER))
    check("no context columns — every printed column has a source",
          ari.CONTEXT_COLUMNS == () and cmp_arc.CONTEXT_FIELDS == ())
    check("Int St Eff-Date is sourced from the intersection row's "
          "InventoryItemStartDate (it was nearly written off as sourceless)",
          "InventoryItemStartDate" in ari.PROJECTION["Int St Eff-Date"][0])
    # The label/position shift is the whole reason this build writes by
    # position: position 9 is LABELLED 'INT Type' but holds the eff-date.
    check("the build writes the export's shifted layout, not the labels",
          ari._POS["INT Eff-Date"] == 9 and ari.HEADER[9] == "INT Type"
          and ari._POS["INT Type"] == 10 and ari.HEADER[10] == "INT Eff-Date")
    check("...and that is the same position the comparator reads by",
          cid._TSMIS_POS["INT Type"] == ari._POS["INT Type"]
          and cid._TSMIS_POS["Int St Eff-Date"] == ari._POS["Int St Eff-Date"]
          and cid._TSMIS_POS["ML Num Lanes"] == ari._POS["ML N/L"])

    print("\nthe row universe — a shell and a retired row are not records:")
    legs = []
    legs += [leg("a1", "1", "Major"), leg("a2", "1", "Minor",
                                          Number_Thru_Lanes=4,
                                          Signal_Arm_Ind="Yes",
                                          N_Distance=150)]
    build_library(
        lib,
        inx_rows=[inx("1", 0.204),
                  # a shell: the IM row exists but carries no inventory block
                  {"INTERSECTION_ID": "2"},
                  # retired
                  inx("3", 0.900, LRS_DATE_RETIRE=D(2024, 1, 1)),
                  # ended inventory
                  inx("4", 1.100, InventoryItemEndDate=D(2025, 1, 1))],
        legs=legs,
        hg_spans=[span(0, 1, "D- Divided Highway", "Highway_Group")],
        city_spans=[span(0, 1, "ANAHEIM", "City_Code")],
        pop_spans=[span(0, 1, "U- Urbanized", "Population_Code")])
    res = ari.consolidate(out_path=out, asof=asof, lib_root=lib,
                          confirm_overwrite=lambda p: True)
    check("build succeeds", res.status == "ok")
    check("...and reports COMPLETE", res.completion == outcome.COMPLETE)
    header, rows = read_out(out)
    check("exactly one record survives the universe rule", len(rows) == 1)
    facts = ari.report_facts(out)
    check("the shell is COUNTED as skipped, not silently dropped",
          facts.get("shells") == "1")
    check("the two retired/ended rows are counted too", facts.get("retired") == "2")
    check("the as-of date is recorded on the marker sheet",
          facts.get("asof") == asof)

    r = rows[0]
    print("\nthe leg rule — Major is the mainline, Minor is the cross street:")
    check("ML Num Lanes comes from the Major leg", cell(r, "ML N/L") == "2")
    check("CS Num Lanes comes from the Minor leg", cell(r, "Inter N") == "4")
    check("ML mast arm comes from the Major leg", cell(r, "ML S/M") == "N")
    check("CS mast arm comes from the Minor leg", cell(r, "Inter S") == "Y")
    check("Main Line Lgth is the Major leg's distance",
          cell(r, "Main Line Lgth") == "250")
    check("Xing Line Lgth is the Minor leg's distance",
          cell(r, "Xing Line Lgth") == "150")

    print("\nthe printed forms:")
    check("Post Mile is zero-padded to 3 decimals",
          cell(r, "Post Mile") == "000.204")
    check("dates print YY-MM-DD", cell(r, "Date of Record") == "73-10-19")
    check("Int St Eff-Date prints the inventory start date",
          cell(r, "Int St Eff-Date") == "23-01-01")
    check("coded domains print the bare code", cell(r, "INT Type") == "T"
          and cell(r, "Ctrl Type") == "B")
    check("Yes/No indicators print Y/N", cell(r, "Light T/Y") == "Y")
    check("the description collapses the layer's padding runs",
          cell(r, "Description") == "MAIN ST LT")
    check("Location is district + county code + route",
          cell(r, "Location") == "12 ORA 001")
    check("the overlay columns are sampled at the postmile",
          cell(r, "H/G") == "D" and cell(r, "R/U") == "U")
    check("...and the City layer's NAME becomes the TASAS city code",
          cell(r, "City Code") == "ANA")

    print("\nthe leg-disagreement rule — the most frequent value is printed:")
    legs2 = [leg("b1", "1", "Major"), leg("b2", "1", "Major"),
             leg("b3", "1", "Major", Number_Thru_Lanes=9),
             leg("b4", "1", "Minor")]
    build_library(lib, inx_rows=[inx("1", 0.204)], legs=legs2,
                  hg_spans=[span(0, 1, "D- Divided Highway", "Highway_Group")],
                  city_spans=[span(0, 1, "ANAHEIM", "City_Code")],
                  pop_spans=[span(0, 1, "U- Urbanized", "Population_Code")])
    res = ari.consolidate(out_path=out, asof=asof, lib_root=lib,
                          confirm_overwrite=lambda p: True)
    _h, rows = read_out(out)
    check("two legs saying 2 outvote one saying 9",
          cell(rows[0], "ML N/L") == "2")

    print("\nthe overlay carry rule — continuous attributes carry, "
          "containment does not:")
    # Every layer has a HOLE over the intersection's postmile (spans stop at
    # 0.100 and resume at 0.500; the intersection sits at 0.204).
    hole = lambda attr, v1, v2: [span(0, 0.100, v1, attr),
                                 span(0.500, 1.0, v2, attr)]
    build_library(
        lib, inx_rows=[inx("1", 0.204)],
        legs=[leg("c1", "1", "Major"), leg("c2", "1", "Minor")],
        hg_spans=hole("Highway_Group", "D- Divided Highway", "U- Undivided"),
        city_spans=hole("City_Code", "ANAHEIM", "ORANGE"),
        pop_spans=hole("Population_Code", "U- Urbanized", "R- Rural"))
    res = ari.consolidate(out_path=out, asof=asof, lib_root=lib,
                          confirm_overwrite=lambda p: True)
    _h, rows = read_out(out)
    r = rows[0]
    check("Highway Group carries the previous span across a gap",
          cell(r, "H/G") == "D")
    check("Rural/Urban carries the previous span across a gap",
          cell(r, "R/U") == "U")
    check("City does NOT carry — a gap means outside the city limits",
          not cell(r, "City Code"))
    check("the carry flags are the measured ones",
          ari.OVERLAY_LAYERS["HG"][2] is True
          and ari.OVERLAY_LAYERS["POP"][2] is True
          and ari.OVERLAY_LAYERS["CITY"][2] is False)

    print("\nthe role gates:")
    check("our build is recognised as ours", ari.is_arcgis_report(out))
    check("a swapped pair is refused (TSMIS side is an ArcGIS build)",
          _raises(lambda: cmp_arc._load_pair(out, out), "not a TSMIS export"))
    check("a non-build ArcGIS side is refused",
          _raises(lambda: cmp_arc._load_pair(lib / f"38_{ari.INX_LAYER}.xlsx",
                                             out),
                  "not an ArcGIS-built"))

    print("\nthe missing-layer refusal:")
    stripped = tmp / "partial"
    stripped.mkdir(exist_ok=True)
    for p in lib.glob("*.xlsx"):
        if ari.APP_LAYER not in p.name:
            (stripped / p.name).write_bytes(p.read_bytes())
    res = ari.consolidate(out_path=out, asof=asof, lib_root=stripped,
                          confirm_overwrite=lambda p: True)
    check("a missing layer refuses and NAMES it",
          res.status == "error" and ari.APP_LAYER in res.message)

    print("\nthe registry:")
    check("Intersection Detail is registered",
          arcgis_reports.is_report("intersection_detail"))
    for key in arcgis_reports.KEYS:
        label, build_mod, cmp_mod, exports = arcgis_reports.resolve(key)
        check(f"{key}: the build module has the lane's surface",
              all(hasattr(build_mod, a) for a in
                  ("REPORT_NAME", "OUT_PATH", "OUT_DIR", "HEADER", "consolidate",
                   "is_arcgis_report", "report_facts", "CONTEXT_COLUMNS")))
        check(f"{key}: the comparator has the lane's surface",
              all(hasattr(cmp_mod, a) for a in
                  ("compare", "suggest_name", "CONTEXT_FIELDS", "_load_pair")))
        check(f"{key}: every export consolidator can supply the TSMIS side",
              bool(exports) and all(hasattr(m, "SUBDIR")
                                    and hasattr(m, "out_path_for")
                                    for m in exports))
        check(f"{key}: the build and comparator agree on context columns",
              tuple(build_mod.CONTEXT_COLUMNS) == tuple(cmp_mod.CONTEXT_FIELDS))
        check(f"{key}: label is non-empty", bool(label))

    print()
    if _fail:
        print(f"FAILED ({len(_fail)}): " + "; ".join(_fail))
        return 1
    print("All ArcGIS Intersection Detail checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
