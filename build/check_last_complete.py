"""CMP-AUD-085 (last-complete publication) + CMP-AUD-089 (durable attempt state).

Owner decision 2026-07-21: the canonical consolidation and the comparison cache
represent the LAST COMPLETE generation. A partial refresh keeps last-good — it
is reported and retryable, but never overwrites verified bytes and is never
promoted to canonical. 089: every rebuild attempt (ok/partial/error/cancelled)
must leave durable per-cell state instead of silently reverting to the old
green cell, and worker terminal counts must separate attempted / succeeded /
failed / cancelled.

The 085 repro drives the finding's own scenario through the SHIPPED layers: a
real two-route Highway Sequence store, the real consolidator, and the real
MatrixCompareWorker -> matrix.build_comparison orchestration (the comparison
adapter itself is stubbed at the matrix facade seam — it is not under test).
"""
from __future__ import annotations

import json
import queue
import sys
import tempfile
import threading
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

import artifact_store  # noqa: E402
import cache_envelope  # noqa: E402
import consolidation_meta  # noqa: E402
import matrix  # noqa: E402
import matrix_state  # noqa: E402
import outcome  # noqa: E402
import owned_dir  # noqa: E402
import paths  # noqa: E402
import tsn_library  # noqa: E402
from comparison_contract import ComparisonCounts, ComparisonOutcome  # noqa: E402
from events import ConsolidateResult  # noqa: E402
from gui_worker_matrix import (BaselineMatrixCompareWorker,  # noqa: E402
                               DayMatrixCompareWorker,
                               MatrixCompareWorker)
from openpyxl import Workbook, load_workbook  # noqa: E402

_failures: list[str] = []


def check(name: str, cond: bool, detail: str = "") -> None:
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        if detail:
            print(f"       {detail}")
        _failures.append(name)


def _route_xlsx(path: Path, route: str, rows: int = 3) -> None:
    """A minimal but VALID per-route Highway Sequence export."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Highway Locations"
    ws.append(["County", "Postmile", "Description"])
    for i in range(rows):
        ws.append([f"C{route}", f"{i:07.3f}", f"LOC {route}-{i}"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _broken_xlsx(path: Path) -> None:
    """A workbook MISSING the required sheet — the finding's broken route 002."""
    wb = Workbook()
    wb.active.title = "Wrong Sheet"
    wb.active["A1"] = "not a highway locations export"
    wb.save(path)


def _tsn_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TSN"
    ws.append(["County", "Postmile", "Description"])
    ws.append(["CA", "000.100", "TSN LOC"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


class _StubComparator:
    """Stands in for compare_highway_sequence_tsn at the matrix facade seam.

    Publishes a REAL workbook + typed generation through the production
    commit path so recording/publication behave exactly as shipped."""

    def __init__(self):
        self.calls: list[str] = []
        self.raise_error = False

    def compare(self, tsmis_path, tsn_path, out_path, events=None,
                confirm_overwrite=None, mode="values", commit_guard=None,
                **_kwargs):
        self.calls.append(str(tsmis_path))
        if self.raise_error:
            raise RuntimeError("injected comparator crash (CMP-AUD-089)")

        def produce(path: Path) -> ConsolidateResult:
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison"
            ws["A1"] = "stub"
            wb.save(path)
            typed = ComparisonOutcome(
                status="ok", completion="complete", verdict="match",
                counts=ComparisonCounts(known=True, paired_rows=1),
                pairing_quality="exact")
            return ConsolidateResult(
                status="ok", output_path=str(path), verdict="match",
                completion="complete", skipped_inputs=0, failed_inputs=0,
                comparison_outcome=typed)

        return artifact_store.commit_workbook(
            Path(out_path), produce, expect_sheet="Comparison",
            requested_mode="values",
            confirm_overwrite=confirm_overwrite or (lambda _p: True),
            commit_guard=commit_guard)


def _run_worker(dest: Path, cell: str, tsn_files, force: bool):
    q: queue.Queue = queue.Queue()
    w = MatrixCompareWorker(str(dest), cell,
                            [("highway_sequence", cell, "tsn")], q,
                            threading.Event(), tsn_files=tsn_files,
                            force_consolidate=force)
    w.start()
    w.join(timeout=300)
    items = []
    while True:
        try:
            items.append(q.get_nowait())
        except queue.Empty:
            break
    assert not w.is_alive(), "worker did not finish"
    return items


def _sheet_rows(path: Path) -> int:
    wb = load_workbook(path)
    try:
        return int(wb["Highway Locations"].max_row or 0)
    finally:
        wb.close()


def _terminal(items):
    return next(p for t, p in items if t == "matrix_done")


def _cell_events(items):
    return [p for t, p in items if t == "matrix_cell"]


def main() -> None:
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_lastcomplete_"))
    saved = (paths.TSN_LIBRARY_ROOT, paths.OUTPUT_ROOT, paths.INPUT_ROOT)
    paths.TSN_LIBRARY_ROOT = tmp / "_lib"
    paths.OUTPUT_ROOT = tmp / "_out"
    paths.INPUT_ROOT = tmp / "_in"
    # The comparison adapter is NOT under test — stub it at the sanctioned
    # matrix-facade seam (matrix_build resolves _m.tsn_comparator_for at call
    # time), keeping the store, consolidator, worker, leases, publication and
    # cache recording all real.
    saved_cmp = matrix.tsn_comparator_for
    stub = _StubComparator()

    def _patched(row_key):
        if row_key == "highway_sequence":
            return stub
        return saved_cmp(row_key)

    matrix.tsn_comparator_for = _patched
    try:
        _main_body(tmp, stub)
    finally:
        (paths.TSN_LIBRARY_ROOT, paths.OUTPUT_ROOT, paths.INPUT_ROOT) = saved
        matrix.tsn_comparator_for = saved_cmp


def _main_body(tmp: Path, stub) -> None:
    print("CMP-AUD-085 keep-last-good through the shipped matrix path:")
    dest = tmp / "store"
    cell = "ssor-prod"
    dest.mkdir()
    owned_dir.ensure_owned_dir(dest / cell, kind="store")
    store = dest / cell / "highway_sequence"
    _route_xlsx(store / "highway_sequence_route_001.xlsx", "001")
    _route_xlsx(store / "highway_sequence_route_002.xlsx", "002")
    tsn = tmp / "picked" / "tsn_highway_sequence_normalized.xlsx"
    _tsn_workbook(tsn)
    selection = tsn_library.create_explicit_selection(tsn)
    tsn_files = {"highway_sequence": selection}

    # ---- build 1: complete two-route canonical ----
    items = _run_worker(dest, cell, tsn_files, force=True)
    done = _terminal(items)
    canonical = matrix.consolidated_store_path(dest / cell / "highway_sequence", "highway_sequence")
    record0 = consolidation_meta.read_outcome(canonical)
    check("first build is complete, trusted, and canonical",
          done["errors"] == 0 and canonical.exists()
          and record0 is not None and record0.trusted
          and record0.completion == outcome.COMPLETE,
          repr(done))
    bytes0 = canonical.read_bytes()
    rows0 = _sheet_rows(canonical)
    cache0 = (matrix_state.load_tsn_results(dest)
              .get("highway_sequence|tsn") or {}).get(cell) or {}
    check("first build records a complete comparison",
          cache0.get("completion") == outcome.COMPLETE, repr(cache0))
    calls_after_first = len(stub.calls) if stub else None

    # ---- the finding's repro: route 002 replaced by a broken workbook ----
    _broken_xlsx(store / "highway_sequence_route_002.xlsx")
    items = _run_worker(dest, cell, tsn_files, force=True)
    done = _terminal(items)

    record1 = consolidation_meta.read_outcome(canonical)
    check("085: the complete canonical is NOT overwritten by the partial attempt",
          canonical.read_bytes() == bytes0,
          f"bytes changed: {len(bytes0)} -> {canonical.stat().st_size}")
    check("085: the canonical sidecar still certifies the COMPLETE generation",
          record1 is not None and record1.trusted
          and record1.completion == outcome.COMPLETE,
          f"completion={getattr(record1, 'completion', None)!r}")
    if stub:
        check("085: the comparator is NOT run over partial-attempt bytes",
              len(stub.calls) == calls_after_first,
              f"calls {calls_after_first} -> {len(stub.calls)}")
    cache1 = (matrix_state.load_tsn_results(dest)
              .get("highway_sequence|tsn") or {}).get(cell) or {}
    check("085: the comparison cache still holds the last COMPLETE record",
          cache1.get("completion") == outcome.COMPLETE
          and cache1.get("generation_id") == cache0.get("generation_id"),
          repr(cache1))
    check("085/089: the failed refresh is a reported, non-green attempt",
          done["errors"] >= 1, repr(done))

    # ---- 089: the attempt is DURABLE per-cell state, not transient log text ----
    try:
        attempts = matrix_state.load_attempts(matrix_state.comparisons_common_root(dest))
        att = (attempts.get("highway_sequence|tsn") or {}).get(cell) or {}
        check("089: a durable last-attempt record exists for the cell",
              att.get("status") in ("partial", "error")
              and bool(att.get("reason")),
              repr(att))
    except AttributeError:
        check("089: a durable last-attempt record exists for the cell", False,
              "matrix_state.load_attempts does not exist")

    # ---- 085: the kept attempt is a DISTINCT, unpromoted, self-describing file ----
    attempt_path = matrix._attempt_sibling(canonical)
    att_rec = (consolidation_meta.read_outcome(attempt_path)
               if attempt_path.exists() else None)
    check("085: the partial attempt is published beside the canonical under its own name",
          attempt_path.exists() and attempt_path != canonical,
          str(attempt_path))
    check("085: the retained attempt certifies itself PARTIAL (never canonical)",
          att_rec is not None and att_rec.completion == outcome.PARTIAL,
          f"completion={getattr(att_rec, 'completion', None)!r}")

    # ---- repair: the next COMPLETE build replaces canonical + clears the attempt ----
    _route_xlsx(store / "highway_sequence_route_002.xlsx", "002", rows=5)
    items = _run_worker(dest, cell, tsn_files, force=True)
    done = _terminal(items)
    record2 = consolidation_meta.read_outcome(canonical)
    rows2 = _sheet_rows(canonical)
    check("a repaired COMPLETE rebuild replaces the canonical (new bytes, both routes)",
          done["errors"] == 0 and canonical.read_bytes() != bytes0
          and record2 is not None and record2.completion == outcome.COMPLETE
          and rows2 > rows0,
          f"errors={done['errors']} rows {rows0}->{rows2}")
    check("085: promotion leaves no attempt file or sidecar behind",
          not attempt_path.exists()
          and not consolidation_meta.meta_path(attempt_path).exists(),
          f"attempt={attempt_path.exists()}")
    try:
        attempts = matrix_state.load_attempts(matrix_state.comparisons_common_root(dest))
        att = (attempts.get("highway_sequence|tsn") or {}).get(cell) or {}
        check("089: the successful rebuild supersedes the failed attempt",
              att.get("status") in (None, "ok"), repr(att))
    except AttributeError:
        check("089: the successful rebuild supersedes the failed attempt", False,
              "matrix_state.load_attempts does not exist")

    # ---- first-build partial (NO complete predecessor): today's flagged behavior stays ----
    print("first-build partial keeps today's flagged-amber behavior:")
    dest2 = tmp / "store2"
    dest2.mkdir()
    owned_dir.ensure_owned_dir(dest2 / cell, kind="store")
    store2 = dest2 / cell / "highway_sequence"
    _route_xlsx(store2 / "highway_sequence_route_001.xlsx", "001")
    _broken_xlsx(store2 / "highway_sequence_route_002.xlsx")
    _run_worker(dest2, cell, tsn_files, force=True)
    canonical2 = matrix.consolidated_store_path(dest2 / cell / "highway_sequence", "highway_sequence")
    rec2 = consolidation_meta.read_outcome(canonical2)
    check("a first build with no complete predecessor persists FLAGGED partial",
          canonical2.exists() and rec2 is not None
          and rec2.completion == outcome.PARTIAL,
          f"exists={canonical2.exists()} completion={getattr(rec2, 'completion', None)!r}")

    # ---- 089: comparator crash on a forced rebuild + honest terminal counts ----
    print("CMP-AUD-089 crash/cancel attempt accounting:")
    if stub:
        stub.raise_error = True
        items = _run_worker(dest, cell, tsn_files, force=True)
        done = _terminal(items)
        stub.raise_error = False
        cache_after_crash = (matrix_state.load_tsn_results(dest)
                             .get("highway_sequence|tsn") or {}).get(cell) or {}
        check("089: a crashed rebuild keeps the prior complete cache record",
              cache_after_crash.get("completion") == outcome.COMPLETE,
              repr(cache_after_crash))
        try:
            attempts = matrix_state.load_attempts(matrix_state.comparisons_common_root(dest))
            att = (attempts.get("highway_sequence|tsn") or {}).get(cell) or {}
            check("089: the crash is durable last-attempt state",
                  att.get("status") == "error" and "crash" in str(att.get("reason", "")),
                  repr(att))
        except AttributeError:
            check("089: the crash is durable last-attempt state", False,
                  "matrix_state.load_attempts does not exist")
        counts_ok = (done.get("attempted") == 1 and done.get("succeeded") == 0
                     and done.get("failed") == 1 and done.get("cancelled_cells") == 0)
        check("089: terminal counts separate attempted/succeeded/failed/cancelled",
              counts_ok, repr(done))

    print("085 the canonical resolver after every attempt:")
    state = matrix.consolidated_state(dest / cell / "highway_sequence",
                                      "highway_sequence")
    check("the canonical resolver still points at the COMPLETE workbook",
          state["path"] == str(canonical) and state["completion"] == outcome.COMPLETE,
          repr(state))

    # ---- 089: the overlay is a DURABLE FILE and reaches the rendered snapshot ----
    print("089 the attempt overlay is durable and rendered:")
    overlay = matrix_state.attempts_path(matrix_state.comparisons_common_root(dest))
    envelope = json.loads(overlay.read_text(encoding="utf-8")) if overlay.exists() else {}
    check("the overlay is a versioned envelope on disk (survives a restart)",
          envelope.get("schema_version") == cache_envelope.SCHEMA_VERSION
          and envelope.get("output_identity") == "attempts"
          and isinstance(envelope.get("payload"), dict),
          repr(envelope)[:200])
    snap = matrix.matrix_snapshot(dest, baseline_key="ssor-prod", envs=[cell],
                                  row_modes={"highway_sequence": "tsn"})
    snap_cmp = snap["cells"]["highway_sequence"][cell]["cmp"] or {}
    last = snap_cmp.get("last_attempt") or {}
    check("the snapshot the grid renders carries the cell's last attempt",
          last.get("status") == "error" and bool(last.get("reason")),
          repr(snap_cmp)[:300])
    check("the prior result is NOT erased by the attempt overlay",
          snap_cmp.get("built") is True,
          repr({k: snap_cmp.get(k) for k in ("built", "stale", "reason")}))
    check("an unknown attempt state is refused, never persisted as a badge",
          matrix_state.record_attempt(
              matrix_state.comparisons_common_root(dest),
              "highway_sequence|tsn", cell, "bogus") is False)

    # ---- 089: cancellation DURING the one cell is counted as cancelled, not done ----
    if stub:
        cancel = threading.Event()
        q: queue.Queue = queue.Queue()

        def _cancelling_compare(*a, **k):
            cancel.set()
            raise RuntimeError("cancelled mid-cell")

        real = stub.compare
        stub.compare = _cancelling_compare  # type: ignore[method-assign]
        try:
            w = MatrixCompareWorker(str(dest), cell,
                                    [("highway_sequence", cell, "tsn")], q,
                                    cancel, tsn_files=tsn_files,
                                    force_consolidate=False)
            w.start()
            w.join(timeout=300)
        finally:
            stub.compare = real  # type: ignore[method-assign]
        items = []
        while True:
            try:
                items.append(q.get_nowait())
            except queue.Empty:
                break
        done = _terminal(items)
        check("089: a cancelled cell is reported cancelled, never a bare success count",
              done.get("cancelled") is True
              and done.get("succeeded", 1) == 0
              and done.get("cancelled_cells") == 1
              and done.get("failed") == 0,
              repr(done))
        att = (matrix_state.load_attempts(
            matrix_state.comparisons_common_root(dest))
            .get("highway_sequence|tsn") or {}).get(cell) or {}
        check("089: the cancelled cell's durable state says cancelled, not failed",
              att.get("status") == "cancelled", repr(att))

    # ---- 089: all three compare workers report the same honest terminal shape ---- #
    print("089 terminal-count shape across all three compare workers:")
    for label, worker in (
            ("Everything", MatrixCompareWorker(str(dest), cell, [], queue.Queue(),
                                               threading.Event())),
            ("by-day", DayMatrixCompareWorker("ssor-prod", [], str(dest),
                                              queue.Queue(), threading.Event())),
            ("vs-baseline", BaselineMatrixCompareWorker(
                "ssor-prod", [], "base", str(dest), queue.Queue(),
                threading.Event()))):
        worker.start()
        worker.join(timeout=60)
        items = []
        while True:
            try:
                items.append(worker.q.get_nowait())
            except queue.Empty:
                break
        payload = _terminal(items)
        check(f"089: the {label} matrix reports attempted/succeeded/failed/cancelled",
              all(isinstance(payload.get(k), int) for k in
                  ("attempted", "succeeded", "failed", "cancelled_cells",
                   "partial_cells")),
              repr(payload))

    # ---- 089: a newer artifact supersedes an older failed attempt --------------- #
    root = matrix_state.comparisons_common_root(dest)
    matrix_state.record_attempt(root, "highway_sequence|tsn", cell, "error",
                                reason="stale attempt", at=1.0)
    fresh = matrix_state._last_attempt_for(
        matrix_state.load_attempts(root), "highway_sequence|tsn", cell,
        {"mtime": time.time()})
    check("089: an attempt older than the rebuilt artifact stops being shown",
          fresh is None, repr(fresh))


if __name__ == "__main__":
    print("CMP-AUD-085/089 last-complete publication + durable attempt state:")
    main()
    if _failures:
        print(f"\n{len(_failures)} check(s) FAILED")
        raise SystemExit(1)
    print("all good")
