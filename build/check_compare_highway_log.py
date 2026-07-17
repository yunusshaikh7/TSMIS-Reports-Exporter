"""Golden check for the Highway Log / PDF-flavor FILE comparators (v0.19.0 R1).

These were the last three comparators still carrying their own hand-rolled
compare() skeleton; R1 moved them onto compare_tsn_common.run_files_compare.
This check locks the migrated surface end-to-end on synthetic fixtures:

  * compare_highway_log.compare — per-route AND consolidated pairs produce the
    approved sheet set with the expected diff counted; the shape-mismatch pair
    and a missing file return the exact user-facing errors.
  * compare_highway_log_pdf.TSMIS_PDF_VS_TSN / TSMIS_PDF_VS_EXCEL and
    compare_intersection_detail_pdf.* — side labels reach the data sheets, and
    compare() runs the same skeleton (dynamic route-ness for HL, static for ID).
  * the shared helpers (suggest_route_name, row_has_data,
    load_consolidated_rows) behave exactly like the inline copies they replaced.

Stdlib + openpyxl; no browser, no network. Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_highway_log.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from openpyxl import Workbook, load_workbook

import compare_highway_log as hl
import compare_highway_log_pdf as hlp
import compare_intersection_detail_pdf as idp
import compare_intersection_detail_tsn as idt
import compare_tsn_common as ctc
import highway_log_columns as hlc
from paths import today_str

_fail = []


def check(name, cond, detail=""):
    if cond:
        print(f"  ok: {name}")
    else:
        print(f"FAIL: {name}" + (f"\n      {detail}" if detail else ""))
        _fail.append(name)


def _write_hl(path, rows, consolidated=False, marker=False, pdf_marker=False):
    """`marker=True` stamps the v5 "TSN Normalization" sheet — a TSN-side
    fixture must carry it since the CMP-AUD-157/045-HL loader gate.
    `pdf_marker=True` stamps the CMP-AUD-066 PDF-conversion marker — a
    "TSMIS (PDF)"-side fixture must carry it since the role-provenance gate."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Highway Log"
    header = (["Route"] + list(hlc.HEADER)) if consolidated else list(hlc.HEADER)
    ws.append(header)
    n = len(header)
    for r in rows:
        ws.append(list(r)[:n] + [None] * max(0, n - len(r)))
    if marker:
        import consolidate_tsn_highway_log as tsn_hl
        tsn_hl._write_marker_sheet(wb)
    if pdf_marker:
        from pdf_table_lib import write_pdf_source_marker
        write_pdf_source_marker(wb)
    wb.save(path)
    wb.close()


def _hl_row(loc, desc):
    r = [None] * 31
    r[0], r[1], r[2] = loc, "R1", desc
    return r


def _sheets(path):
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def test_hl_compare():
    print("compare_highway_log on the shared skeleton:")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_hlcmp_"))
    a, b = tmp / "TSMIS Route 005.xlsx", tmp / "TSN Route 005.xlsx"
    _write_hl(a, [_hl_row("1.000", "SAME"), _hl_row("2.000", "OLD")])
    _write_hl(b, [_hl_row("1.000", "SAME"), _hl_row("2.000", "NEW")], marker=True)
    out = tmp / "out.xlsx"
    res = hl.compare(str(a), str(b), str(out), confirm_overwrite=lambda p: True,
                     mode="values")
    check("per-route compare runs ok", res.status == "ok", res.message)
    names = _sheets(res.output_path)
    check("approved sheet set present",
          all(s in names for s in ("Summary", "Comparison", "Only in TSMIS",
                                   "Only in TSN", "TSMIS", "TSN")), str(names))
    check("Legend AND the claims Notes sheet are both present (CMP-AUD-157)",
          "Legend" in names and "Notes" in names, str(names))

    # the marker gate (CMP-AUD-157/045-HL): a pre-v5 TSN side refuses
    b_old = tmp / "TSN old.xlsx"
    _write_hl(b_old, [_hl_row("1.000", "SAME")])
    res_old = hl.compare(str(a), str(b_old), str(tmp / "out_old.xlsx"),
                         confirm_overwrite=lambda p: True, mode="values")
    check("a pre-v5 (markerless) TSN side refuses with the rebuild hint",
          res_old.status == "error"
          and "older TSN Highway Log converter" in res_old.message
          and "rebuild the TSN library" in res_old.message, res_old.message)

    ca, cb = tmp / "TSMIS Consolidated.xlsx", tmp / "TSN Consolidated.xlsx"
    _write_hl(ca, [["001"] + _hl_row("1.000", "X")], consolidated=True)
    _write_hl(cb, [["001"] + _hl_row("1.000", "Y")], consolidated=True,
              marker=True)
    res2 = hl.compare(str(ca), str(cb), str(tmp / "out2.xlsx"),
                      confirm_overwrite=lambda p: True, mode="values")
    check("consolidated compare runs ok (dynamic route-ness)",
          res2.status == "ok", res2.message)

    res3 = hl.compare(str(a), str(cb), str(tmp / "out3.xlsx"))
    check("shape mismatch -> the exact user-facing error",
          res3.status == "error"
          and "different shapes" in res3.message
          and "consolidated workbook (has a Route column)" in res3.message
          and "per-route" in res3.message, res3.message)

    res4 = hl.compare(str(tmp / "nope.xlsx"), str(b), str(tmp / "out4.xlsx"))
    check("missing TSMIS file names the TSMIS side",
          res4.status == "error"
          and res4.message.startswith("The TSMIS file doesn't exist:"), res4.message)

    # CMP-AUD-049 (direct-compare half): a per-route pair must carry MATCHING
    # filename route tokens — per-route workbooks have no Route column, so
    # the filename is the only identity a direct comparison can verify.
    a2 = tmp / "TSMIS Route 002.xlsx"
    _write_hl(a2, [_hl_row("1.000", "SAME")])
    res5 = hl.compare(str(a2), str(b), str(tmp / "out5.xlsx"),
                      confirm_overwrite=lambda p: True, mode="values")
    check("per-route pair naming DIFFERENT routes refuses (CMP-AUD-049)",
          res5.status == "error" and "different routes" in res5.message
          and "002" in res5.message and "005" in res5.message, res5.message)
    a3 = tmp / "tsmis export.xlsx"
    _write_hl(a3, [_hl_row("1.000", "SAME")])
    res6 = hl.compare(str(a3), str(b), str(tmp / "out6.xlsx"),
                      confirm_overwrite=lambda p: True, mode="values")
    check("a per-route file without a route token refuses with guidance",
          res6.status == "error" and "route token" in res6.message
          and "consolidated" in res6.message, res6.message)


def test_pdf_flavors():
    print("the PDF-sourced flavors ride the same skeleton:")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_hlpdf_"))
    a, b = tmp / "pdf route 7.xlsx", tmp / "tsn route 7.xlsx"
    _write_hl(a, [_hl_row("1.000", "P")], pdf_marker=True)
    _write_hl(b, [_hl_row("1.000", "Q")], marker=True)
    res = hlp.TSMIS_PDF_VS_TSN.compare(str(a), str(b), str(tmp / "o.xlsx"),
                                       confirm_overwrite=lambda p: True, mode="values")
    check("HL PDF-vs-TSN runs ok", res.status == "ok", res.message)
    names = _sheets(res.output_path)
    check("HL PDF side labels reach the data sheets",
          "TSMIS (PDF)" in names and "TSN (PDF)" in names, str(names))
    b_old = tmp / "tsn_old route 7.xlsx"
    _write_hl(b_old, [_hl_row("1.000", "Q")])
    res_gate = hlp.TSMIS_PDF_VS_TSN.compare(str(a), str(b_old),
                                            str(tmp / "og.xlsx"),
                                            confirm_overwrite=lambda p: True,
                                            mode="values")
    check("HL PDF-vs-TSN refuses a pre-v5 TSN side",
          res_gate.status == "error"
          and "older TSN Highway Log converter" in res_gate.message,
          res_gate.message)
    res_ex = hlp.TSMIS_PDF_VS_EXCEL.compare(str(a), str(b_old),
                                            str(tmp / "oe.xlsx"),
                                            confirm_overwrite=lambda p: True,
                                            mode="values")
    check("HL PDF-vs-Excel has no TSN side and does NOT gate on the TSN marker "
          "(its Excel side stays unmarked by design)",
          res_ex.status == "ok", res_ex.message)
    b5 = tmp / "excel route 5.xlsx"
    _write_hl(b5, [_hl_row("1.000", "Q")])
    res_mm = hlp.TSMIS_PDF_VS_EXCEL.compare(str(a), str(b5),
                                            str(tmp / "omm.xlsx"),
                                            confirm_overwrite=lambda p: True,
                                            mode="values")
    check("the PDF flavors inherit the per-route identity rule (CMP-AUD-049)",
          res_mm.status == "error" and "different routes" in res_mm.message,
          res_mm.message)
    resm = hlp.TSMIS_PDF_VS_EXCEL.compare(str(tmp / "nope.xlsx"), str(b),
                                          str(tmp / "o2.xlsx"))
    check("HL PDF missing-file error names the flavor's OWN side label",
          resm.status == "error"
          and resm.message.startswith("The TSMIS (PDF) file doesn't exist:"),
          resm.message)

    # Intersection Detail flavors: consolidated TSMIS both sides (vs-Excel), the
    # July-2026 35-column shape (the loader's header gate demands its tail column).
    def id_row(route, pm, desc):
        r = [None] * 36
        r[0], r[1], r[2], r[4], r[5] = route, "R", pm, f"12 ORA {route}", "73-10-19"
        r[21] = desc
        return r

    ia, ib = tmp / "id_pdf.xlsx", tmp / "id_xls.xlsx"
    for p, d in ((ia, "MAIN ST"), (ib, "MAIN STREET")):
        wb = Workbook()
        ws = wb.active
        ws.title = idt.TSMIS_SHEET
        ws.append(list(idt._TSMIS_HEADER))       # CMP-AUD-034 exact header
        ws.append(id_row("001", "1.000", d))
        if p is ia:
            # The "TSMIS (PDF)" side must carry the CMP-AUD-066 marker.
            from pdf_table_lib import write_pdf_source_marker
            write_pdf_source_marker(wb)
        wb.save(p)
        wb.close()
    res2 = idp.TSMIS_PDF_VS_EXCEL.compare(str(ia), str(ib), str(tmp / "o3.xlsx"),
                                          confirm_overwrite=lambda p: True,
                                          mode="values")
    check("ID PDF-vs-Excel runs ok (static route-ness)", res2.status == "ok",
          res2.message)
    names2 = _sheets(res2.output_path)
    check("ID PDF-vs-Excel labels + dropped Notes sheet",
          "TSMIS (PDF)" in names2 and "TSMIS (Excel)" in names2
          and "Notes" not in names2, str(names2))


def test_shared_helpers():
    print("the shared helpers match the inline copies they replaced:")
    d = today_str()
    check("suggest_route_name: route token (zeros stripped)",
          ctc.suggest_route_name("TSMIS Route 005 x.xlsx", "FB", "TAG")
          == f"TAG_Route5_Comparison {d}.xlsx")
    check("suggest_route_name: consolidated",
          ctc.suggest_route_name("Consolidated x.xlsx", "FB", "TAG")
          == f"TAG_Consolidated_Comparison {d}.xlsx")
    check("suggest_route_name: fallback",
          ctc.suggest_route_name("misc.xlsx", "FB", "TAG")
          == f"TAG_FB_Comparison {d}.xlsx")
    check("hl.suggest_name unchanged shape",
          hl.suggest_name("Route 12.xlsx") == f"TSMIS_vs_TSN_Route12_Comparison {d}.xlsx")

    check("row_has_data: whitespace-only row is empty",
          not ctc.row_has_data(["", None, "  "]) and ctc.row_has_data([None, 0]))

    tmp = Path(tempfile.mkdtemp(prefix="tsmis_lcr_"))
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["NotRoute", "x"])
    wb.save(tmp / "bad.xlsx")
    wb.close()
    try:
        ctc.load_consolidated_rows(tmp / "bad.xlsx", "S",
                                   missing_sheet_hint="h.", bad_header_msg="is bad.")
        check("load_consolidated_rows: bad header raises", False)
    except ValueError as e:
        check("load_consolidated_rows: bad header raises the composed message",
              str(e) == "bad.xlsx is bad.", str(e))
    try:
        ctc.load_consolidated_rows(tmp / "bad.xlsx", "Missing",
                                   missing_sheet_hint="pick the right workbook.",
                                   bad_header_msg="is bad.")
        check("load_consolidated_rows: missing sheet raises", False)
    except ValueError as e:
        check("load_consolidated_rows: missing sheet raises the composed message",
              str(e) == "bad.xlsx has no 'Missing' sheet — pick the right workbook.",
              str(e))


def main():
    test_hl_compare()
    test_pdf_flavors()
    test_shared_helpers()
    if _fail:
        print(f"\n{len(_fail)} check(s) FAILED")
        return 1
    print("\nall good")
    return 0


if __name__ == "__main__":
    sys.exit(main())
