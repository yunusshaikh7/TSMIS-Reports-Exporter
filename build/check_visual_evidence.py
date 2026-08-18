"""Golden check for the visual-evidence generator (scripts/visual_evidence.py +
scripts/evidence_highway_detail.py) — the render-free logic layer.

Locks: the row registry + the TSMIS-PDF/TSN-PDF source resolution and the
examples clamp; the caller-side gate (matrix_build.evidence_opts_for); the
sibling artifact naming (the "(formulas).xlsx" family); the adapter's LOCKSTEP
pins against the Highway Detail PDF consolidator (window counts, the postmile /
date-token regex behavior its mirrored walk relies on); the field→TSN-print
group map (complete, RB half mirrored) and the two-line TASAS regexes on
realistic print lines (prefix/roadbed/equation/optional-city/empty-description);
the span→x-box math including the empty-optional-group case; the verification
projections (PS derived, NA fold via the comparator's own projection); the
unique-key diff enumeration with the district/county sidecar; and the TSN
loader's sidecar contract (tsn_rows_with_dcr row-identical to the locked
tsn_rows_from_raw; the normalized sheet appending exactly the sidecar columns;
load_sides reading it back and refusing a sidecar-less legacy library with the
rebuild hint). Rendering itself is exercised by the frozen self-test's
render-stack step (scripts/self_test.py) — no rasterizing here; CI-safe.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_visual_evidence.py
"""
import inspect
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import _checklib
import compare_highway_detail_tsn as cht
import compare_highway_log as chl_cmp
import compare_highway_sequence_tsn as chsl_cmp
import compare_ramp_detail_pdf as crdp
import compare_ramp_detail_tsn as crd_cmp
import consolidate_tsmis_highway_detail_pdf as chd
import consolidate_tsmis_highway_sequence_pdf as chslp
import consolidate_tsmis_ramp_detail_pdf as crdpdf
import consolidate_tsn_highway_log as ctnl
import consolidate_tsn_highway_sequence as ctnsl
import evidence_highway_detail as ehd
import evidence_highway_log as ehl
import evidence_highway_sequence as ehsl
import evidence_ramp_detail as erd
import highway_detail_columns as hdc
import highway_log_columns as hlc
import matrix_build
import tsn_library
import tsn_load_highway_detail as tlh
import visual_evidence as ve
import artifact_store
import consolidation_meta
from comparison_contract import ArtifactGeneration, ComparisonCounts, ComparisonOutcome
from events import ConsolidateResult
from openpyxl import Workbook, load_workbook

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


# --------------------------------------------------------------------------- #
# Stated FIRST, as an assertion rather than an import-time crash: run against a
# runtime without the print-crop amendment, this file used to die on the first
# new call and print nothing, so it could never demonstrate the defect it
# exists to catch. Naming the contract makes the base run say what is missing.
print("the print-crop contract this file depends on (owner amendment 2026-08-05)")
check("the read set is captured in labelled per-side buckets, so a manifest "
      "names the documents each side's crops were read from",
      "buckets" in inspect.signature(ve._snapshot_read_set).parameters)
check("evidence is refused outside the PDF-edition families at the engine "
      "boundary (capable names only _pdf rows; FLAVOR_SELF not in FLAVORS)",
      not ve.capable("highway_log") and not ve.self_capable("highway_log_pdf")
      and ve.FLAVOR_SELF not in ve.FLAVORS)
check("the vs-TSN lane locates BOTH prints (_locate_tsmis_sources + the "
      "adapter tsn locate loop), never a workbook panel",
      hasattr(ve, "_locate_tsmis_sources")
      and "tsn_ctx" in inspect.signature(ve._try_example).parameters)
check("the per-route print is resolved by contract (find_route_print)",
      hasattr(ve, "find_route_print"))
check("a print target box is engine-checked against the record's own printed "
      "lines and width (_box_within_record)",
      hasattr(ve, "_box_within_record"))

# --------------------------------------------------------------------------- #
print("registry + sources + clamp")
check("rows: the five PDF-edition rows, nothing else (the 2026-08-05 ruling — "
      "the Excel rows stay out; HD-PDF JOINED in v0.37.0 when the vendor's "
      "official release lifted the pre-release freeze)",
      ve.rows() == ["highway_detail_pdf", "highway_log_pdf", "highway_sequence_pdf",
                    "intersection_detail_pdf", "ramp_detail_pdf"])
check("capable() matches rows(), and refuses every non-_pdf row",
      all(ve.capable(r) for r in ve.rows())
      and not ve.capable("ramp_summary") and not ve.capable("highway_log")
      and not ve.capable("highway_detail")          # the EXCEL row stays out
      and not ve.capable("intersection_detail")
      and not ve.capable("ramp_detail"))
# HD-PDF is vs-TSN ONLY: its adapter has locate_tsn/tsn_value/tsn_box but no
# env hooks, so the ENV lane must still refuse it (env_capable checks the hooks).
check("highway_detail_pdf is vs-TSN capable but NOT env capable",
      ve.capable("highway_detail_pdf") and not ve.env_capable("highway_detail_pdf"))
check("the ENV lane stays at the original four rows",
      sorted(ve._ENV_ADAPTER_MODULES) == ["highway_log_pdf", "highway_sequence_pdf",
                                          "intersection_detail_pdf", "ramp_detail_pdf"])
check("the self lane is retired everywhere (self_capable False for every row)",
      not any(ve.self_capable(r) for r in ve.rows()))
check("TSMIS visuals come from each report's (PDF)-edition export subdir",
      ve.pdf_subdir_for("intersection_detail_pdf") == "intersection_detail_pdf"
      and ve.pdf_subdir_for("highway_log_pdf") == "highway_log_pdf"
      and ve.pdf_subdir_for("highway_sequence_pdf") == "highway_sequence_pdf"
      and ve.pdf_subdir_for("ramp_detail_pdf") == "ramp_detail_pdf"
      and set(ve.TSMIS_PDF_SUBDIR) == set(ve.rows())
      and set(ve.TSN_PDF_REPORT) == set(ve.rows()))
check("TSN prints live in each report's library pdf folder — except the Highway "
      "Log and Highway Sequence, whose district prints ARE the library's raw "
      "inputs (no duplicate drop)",
      str(ve.tsn_pdf_dir("intersection_detail_pdf")).replace("\\", "/")
      .endswith("tsn_library/intersection_detail/pdf")
      and str(ve.tsn_pdf_dir("ramp_detail_pdf")).replace("\\", "/")
      .endswith("tsn_library/ramp_detail/pdf")
      and str(ve.tsn_pdf_dir("highway_log_pdf")).replace("\\", "/")
      .endswith("tsn_library/highway_log/raw")
      and str(ve.tsn_pdf_dir("highway_sequence_pdf")).replace("\\", "/")
      .endswith("tsn_library/highway_sequence/raw"))
check("clamp: default/garbage/low/high",
      (ve.clamp_examples(None), ve.clamp_examples("x"), ve.clamp_examples(0),
       ve.clamp_examples(99), ve.clamp_examples("7"))
      == (2, 2, 1, 10, 7))
wbp, imgp = ve.sibling_paths(Path(r"C:\x\comparisons\hd vs tsn.xlsx"))
check("sibling naming: '(evidence).xlsx' + '(evidence images)' folder",
      wbp.name == "hd vs tsn (evidence).xlsx"
      and imgp.name == "hd vs tsn (evidence images)")
_alias_tmp = Path(tempfile.mkdtemp(prefix="evidence_alias_guard_"))
try:
    _alias_cmp = _alias_tmp / "comparison.xlsx"
    _alias_wb, _alias_img = ve.sibling_paths(_alias_cmp)
    _alias_wb.write_bytes(b"selected source")
    try:
        ve._safe_sibling_paths(_alias_cmp, (_alias_wb,))
        _wb_alias_rejected = False
    except ValueError:
        _wb_alias_rejected = True
    check("the derived evidence workbook cannot alias a comparison source",
          _wb_alias_rejected and _alias_wb.read_bytes() == b"selected source")
    _alias_wb.unlink()
    _alias_img.mkdir()
    try:
        ve._safe_sibling_paths(_alias_cmp, (_alias_img,))
        _dir_alias_rejected = False
    except ValueError:
        _dir_alias_rejected = True
    check("the derived evidence image folder cannot alias a source directory",
          _dir_alias_rejected and _alias_img.is_dir())

    # Stable identity must survive a rename+decoy race across the image-folder
    # transaction, not merely compare the two current pathnames.
    _swap_source = _alias_tmp / "swap-source"
    _swap_source.mkdir()
    (_swap_source / "selected.txt").write_bytes(b"selected source directory")
    _swap_tmp = _alias_tmp / "rendered.tmp"
    _swap_tmp.mkdir()
    (_swap_tmp / "new.txt").write_bytes(b"new evidence")
    _swap_target = _alias_tmp / "evidence-images"
    _captured = ve.artifact_store.capture_source_identities((_swap_source,))
    _real_guard = ve.artifact_store.ensure_outputs_do_not_alias_sources
    _guard_calls = [0]

    def _swap_then_guard(destinations, sources, **kwargs):
        _guard_calls[0] += 1
        if _guard_calls[0] == 2:
            os.replace(_swap_source, _swap_target)
            _swap_source.mkdir()                    # same-name decoy
        return _real_guard(destinations, sources, **kwargs)

    ve.artifact_store.ensure_outputs_do_not_alias_sources = _swap_then_guard
    try:
        try:
            ve._swap_dir(
                _swap_tmp, _swap_target, source_paths=(_swap_source,),
                captured_sources=_captured)
            _swap_rejected = False
        except ValueError:
            _swap_rejected = True
    finally:
        ve.artifact_store.ensure_outputs_do_not_alias_sources = _real_guard
    check("a renamed evidence source plus decoy is rejected during folder swap",
          _swap_rejected)
    check("...the originally selected directory survives at its moved path",
          (_swap_target / "selected.txt").read_bytes()
          == b"selected source directory")

    # A guard failure after canonical target -> quarantine used to escape the
    # OSError-only handler and leave the canonical image directory missing.
    _rollback_target = _alias_tmp / "rollback-images"
    _rollback_target.mkdir()
    (_rollback_target / "prior.txt").write_bytes(b"prior image set")
    _rollback_tmp = _alias_tmp / "rollback-rendered"
    _rollback_tmp.mkdir()
    (_rollback_tmp / "new.txt").write_bytes(b"new image set")
    _source_checks = [0]

    def _fail_after_quarantine():
        _source_checks[0] += 1
        if _source_checks[0] == 2:
            raise ValueError("source set changed after quarantine")

    try:
        ve._swap_dir(_rollback_tmp, _rollback_target,
                     source_set_check=_fail_after_quarantine)
        _rollback_rejected = False
    except ValueError:
        _rollback_rejected = True
    check("ValueError after target quarantine restores the canonical directory",
          _rollback_rejected
          and (_rollback_target / "prior.txt").read_bytes() == b"prior image set"
          and (_rollback_tmp / "new.txt").read_bytes() == b"new image set")

    # The same rollback must run when the canonical publish first raises OSError
    # and a guard then fails before the alternate move (the former pre-alt gap).
    _late_target = _alias_tmp / "late-guard-images"
    _late_target.mkdir()
    (_late_target / "prior.txt").write_bytes(b"late prior image set")
    _late_tmp = _alias_tmp / "late-guard-rendered"
    _late_tmp.mkdir()
    (_late_tmp / "new.txt").write_bytes(b"late new image set")
    _late_checks = [0]
    _real_replace = ve.os.replace

    def _late_guard():
        _late_checks[0] += 1
        if _late_checks[0] == 3:       # fallback guard, after target -> old
            raise ValueError("source set changed before alternate")

    def _fail_canonical_publish(src, dst):
        if Path(src) == _late_tmp and Path(dst) == _late_target:
            raise PermissionError("simulated locked image destination")
        return _real_replace(src, dst)

    ve.os.replace = _fail_canonical_publish
    try:
        try:
            ve._swap_dir(_late_tmp, _late_target,
                         source_set_check=_late_guard)
            _late_rejected = False
        except ValueError:
            _late_rejected = True
    finally:
        ve.os.replace = _real_replace
    check("fallback guard failure also restores the quarantined prior set",
          _late_rejected
          and (_late_target / "prior.txt").read_bytes() == b"late prior image set"
          and (_late_tmp / "new.txt").read_bytes() == b"late new image set")

    # Fixed .old/.new names may already contain unrelated material. The swap
    # now uses random per-operation names and must never delete those sentinels.
    _foreign_target = _alias_tmp / "foreign-images"
    _foreign_target.mkdir()
    (_foreign_target / "prior.txt").write_bytes(b"prior")
    _foreign_tmp = _alias_tmp / "foreign-rendered"
    _foreign_tmp.mkdir()
    (_foreign_tmp / "new.txt").write_bytes(b"published")
    _fixed_old = _alias_tmp / "foreign-images.old"
    _fixed_new = _alias_tmp / "foreign-images.new"
    _fixed_old.mkdir(); (_fixed_old / "sentinel.txt").write_bytes(b"foreign old")
    _fixed_new.mkdir(); (_fixed_new / "sentinel.txt").write_bytes(b"foreign new")
    ve._swap_dir(_foreign_tmp, _foreign_target)
    check("foreign fixed .old/.new directories survive a successful swap",
          (_fixed_old / "sentinel.txt").read_bytes() == b"foreign old"
          and (_fixed_new / "sentinel.txt").read_bytes() == b"foreign new"
          and (_foreign_target / "new.txt").read_bytes() == b"published")

    # Workbook lock fallback is random + exclusively reserved too; a legacy
    # fixed '<stem>.new.xlsx' file is foreign and must remain byte-identical.
    _fallback_wb = _alias_tmp / "fallback evidence.xlsx"
    _fallback_wb.write_bytes(b"locked prior workbook")
    _fixed_wb_alt = _alias_tmp / "fallback evidence.new.xlsx"
    _fixed_wb_alt.write_bytes(b"foreign workbook sentinel")
    _fallback_images = _alias_tmp / "fallback-images"; _fallback_images.mkdir()
    _replace_calls = [0]
    _real_replace = ve.os.replace

    def _lock_only_primary(src, dst):
        if Path(dst) == _fallback_wb and _replace_calls[0] == 0:
            _replace_calls[0] += 1
            raise PermissionError("simulated workbook open in Excel")
        return _real_replace(src, dst)

    ve.os.replace = _lock_only_primary
    try:
        _fallback_note = ve._write_workbook(
            _fallback_wb, _fallback_images, [], {},
            {"report": "Probe", "comparison": "probe.xlsx",
             "examples": 1, "seed": "00000000",
             "tsmis_dir": "A", "tsn_dir": "B"})
    finally:
        ve.os.replace = _real_replace
    _random_alts = list(_alias_tmp.glob("fallback evidence.new-*.xlsx"))
    check("foreign fixed workbook .new sentinel is preserved",
          _fixed_wb_alt.read_bytes() == b"foreign workbook sentinel")
    check("locked workbook diverts only to a disclosed random fallback",
          len(_random_alts) == 1 and _random_alts[0].name in _fallback_note)

    # Matrix/Everything evidence is nested under an exact comparisons lease.
    # The lease must reach the workbook and image transactions as a target-aware
    # predicate, not merely as a one-time worker preflight.
    _leased_root = _alias_tmp / "leased-comparisons"
    _lease = ve.owned_dir.require_owned_dir_lease(
        _leased_root, kind="comparisons")
    _leased_tmp = _leased_root / "fresh-images"
    _leased_tmp.mkdir()
    (_leased_tmp / "new.txt").write_bytes(b"new leased images")
    _leased_target = _leased_root / "evidence-images"
    ve._swap_dir(
        _leased_tmp, _leased_target, commit_guard=_lease.guard,
        tmp_directory_identity=ve.owned_dir.directory_identity(_leased_tmp))
    check("a current comparisons lease authorizes the complete image swap",
          (_leased_target / "new.txt").read_bytes() == b"new leased images")

    _guarded_wb = _leased_root / "guarded evidence.xlsx"
    _guarded_wb.write_bytes(b"prior guarded workbook")
    _guarded_images = _leased_root / "guarded-images"
    _guarded_images.mkdir()
    _guard_paths = []

    def _deny_workbook_publish(path=None, **kwargs):
        if path is None:
            return _lease.is_current()
        path = Path(path)
        _guard_paths.append(path)
        return path != _guarded_wb and _lease.guard(path, **kwargs)

    try:
        ve._write_workbook(
            _guarded_wb, _guarded_images, [], {},
            {"report": "Probe", "comparison": "probe.xlsx",
             "examples": 1, "seed": "00000000",
             "tsmis_dir": "A", "tsn_dir": "B"},
            commit_guard=_deny_workbook_publish)
        _guarded_wb_rejected = False
    except ve.owned_dir.OwnershipError:
        _guarded_wb_rejected = True
    check("a late target-aware workbook guard preserves the prior workbook",
          _guarded_wb_rejected
          and _guarded_wb.read_bytes() == b"prior guarded workbook"
          and _guarded_wb in _guard_paths
          and not list(_leased_root.glob(".guarded evidence.tmp-*.xlsx")))

    # If the fresh temp becomes untrusted after prior images were quarantined,
    # rollback is still allowed through the live lease and restores last-good.
    _rollback_guard_target = _leased_root / "guard-rollback-images"
    _rollback_guard_target.mkdir()
    (_rollback_guard_target / "prior.txt").write_bytes(b"guarded prior")
    _rollback_guard_tmp = _leased_root / "guard-rollback-fresh"
    _rollback_guard_tmp.mkdir()
    (_rollback_guard_tmp / "new.txt").write_bytes(b"guarded new")
    _reject_guard_tmp = [False]
    _real_replace = ve.os.replace

    def _reject_temp_after_quarantine(src, dst):
        result = _real_replace(src, dst)
        if (Path(src) == _rollback_guard_target
                and Path(dst).name.startswith(
                    _rollback_guard_target.name + ".old-")):
            _reject_guard_tmp[0] = True
        return result

    def _selective_guard(path=None, **kwargs):
        if path is None:
            return _lease.is_current()
        if _reject_guard_tmp[0] and Path(path) == _rollback_guard_tmp:
            return False
        return _lease.guard(path, **kwargs)

    ve.os.replace = _reject_temp_after_quarantine
    try:
        try:
            ve._swap_dir(
                _rollback_guard_tmp, _rollback_guard_target,
                commit_guard=_selective_guard,
                tmp_directory_identity=ve.owned_dir.directory_identity(
                    _rollback_guard_tmp))
            _late_lease_rejected = False
        except ve.owned_dir.OwnershipError:
            _late_lease_rejected = True
    finally:
        ve.os.replace = _real_replace
    check("a guard failure after quarantine restores last-good images",
          _late_lease_rejected
          and (_rollback_guard_target / "prior.txt").read_bytes()
          == b"guarded prior"
          and (_rollback_guard_tmp / "new.txt").read_bytes()
          == b"guarded new")

    # Workbook serialization uses an exclusive unpredictable temp handle.  If
    # a selected source is moved onto that temp after serialization but before
    # publication, the post-save identity check must reject it and cleanup must
    # retain (not unlink) the selected object.
    _save_source = _alias_tmp / "save-source.xlsx"
    _save_source.write_bytes(b"selected workbook source")
    _save_prior = _save_source.read_bytes()
    _save_target = _alias_tmp / "evidence.xlsx"
    _save_img_dir = _alias_tmp / "empty-images"; _save_img_dir.mkdir()
    _save_captured = ve.artifact_store.capture_source_identities((_save_source,))
    _real_save = ve.Workbook.save
    _moved_save_temp = [None]

    def _save_then_swap_source(workbook, target):
        _real_save(workbook, target)
        target.flush()
        target.close()
        _moved_save_temp[0] = Path(target.name)
        os.replace(_save_source, _moved_save_temp[0])
        _save_source.write_bytes(b"same-path decoy")

    ve.Workbook.save = _save_then_swap_source
    try:
        try:
            ve._write_workbook(
                _save_target, _save_img_dir, [], {},
                {"report": "Probe", "comparison": "probe.xlsx",
                 "examples": 1, "seed": "00000000",
                 "tsmis_dir": "A", "tsn_dir": "B"},
                source_paths=(_save_source,),
                captured_sources=_save_captured)
            _save_swap_rejected = False
        except ValueError:
            _save_swap_rejected = True
    finally:
        ve.Workbook.save = _real_save
    check("a source moved onto the evidence temp at save-time is rejected",
          _save_swap_rejected and not _save_target.exists())
    check("...source-safe cleanup retains the selected workbook bytes",
          _moved_save_temp[0].read_bytes() == _save_prior)
    _moved_save_temp[0].unlink()

    # CMP-AUD-098 + 112: the READ SET is a private snapshot, and the bytes the
    # images illustrate come from that copy — not from a live path that anything
    # else may still be writing to. Everything below drives the real snapshot.
    _cpdf = _alias_tmp / "content.pdf"
    _cpdf.write_bytes(b"%PDF-1.4 original bytes")
    _c_mtime = _cpdf.stat().st_mtime_ns
    _c_ids_before = ve.artifact_store.canonical_path_identities((_cpdf,))
    _rs = ve._snapshot_read_set([_cpdf], [])
    try:
        _snap = _rs.tsmis_dir / _cpdf.name
        check("the snapshot copies the PDF under its own basename "
              "(the adapters look files up by name)",
              _snap.is_file() and _snap.read_bytes() == _cpdf.read_bytes())
        check("the read set digests the COPY and names the ORIGINAL path",
              len(_rs.members) == 1 and _rs.members[0].name == str(_cpdf)
              and _rs.members[0].sha256
              == _rs.digests[str(_cpdf.resolve())]
              == ve._pdf_content_digests([_snap])[str(_snap.resolve())])
        _rs.ensure_sources_unchanged()                     # unchanged -> no raise

        # The A->B->A swap: digest A, parse B, restore A. A start/end digest of
        # the LIVE file passes and certifies bytes nobody rendered from; reading
        # the snapshot means the swap never reaches the render at all.
        _cpdf.write_bytes(b"%PDF-1.4 SWAPPED! bytes")       # same length (23)
        os.utime(_cpdf, ns=(_c_mtime, _c_mtime))           # restore the mtime
        check("the metadata tripwire alone MISSES a same-size same-mtime swap",
              ve.artifact_store.canonical_path_identities((_cpdf,)) == _c_ids_before)
        check("...but the snapshot still holds the bytes evidence will read",
              _snap.read_bytes() == b"%PDF-1.4 original bytes")
        try:
            _rs.ensure_sources_unchanged()
            _content_rejected = False
        except ValueError:
            _content_rejected = True
        check("a metadata-preserving content change aborts the evidence publish",
              _content_rejected)
        _cpdf.write_bytes(b"%PDF-1.4 original bytes")       # A restored
        os.utime(_cpdf, ns=(_c_mtime, _c_mtime))
        _rs.ensure_sources_unchanged()                      # A->B->A: back in sync
        check("A->B->A leaves the rendered bytes untouched (CMP-AUD-098)",
              _snap.read_bytes() == b"%PDF-1.4 original bytes")
        # a deleted/unreadable source also aborts (provenance no longer current)
        _cpdf.unlink()
        try:
            _rs.ensure_sources_unchanged()
            _gone_rejected = False
        except ValueError:
            _gone_rejected = True
        check("a source PDF that vanished before publish also aborts",
              _gone_rejected)
    finally:
        _rs.discard()
    check("discard removes the snapshot", not _rs.root.exists())

    # HF-05/HF-10: labelled side buckets (the env flavor's two run folders may
    # hold same-named per-route files), original-stat capture for the census
    # binding, and durable member renaming (PCOA-FINAL-003: a manifest must
    # never name a private capture path).
    _sa = _alias_tmp / "side-a"; _sa.mkdir()
    _sb = _alias_tmp / "side-b"; _sb.mkdir()
    (_sa / "route_001.pdf").write_bytes(b"%PDF a-side")
    (_sb / "route_001.pdf").write_bytes(b"%PDF b-side!")
    _rs2 = ve._snapshot_read_set((), (), buckets=(
        ("side_a", [_sa / "route_001.pdf"]), ("side_b", [_sb / "route_001.pdf"])))
    try:
        check("same-named files land in their own side buckets, bytes intact",
              (_rs2.dir_for("side_a") / "route_001.pdf").read_bytes() == b"%PDF a-side"
              and (_rs2.dir_for("side_b") / "route_001.pdf").read_bytes() == b"%PDF b-side!")
        check("original size+mtime are captured for the env census binding",
              _rs2.stats[str((_sa / "route_001.pdf").resolve())]
              == ((_sa / "route_001.pdf").stat().st_size,
                  (_sa / "route_001.pdf").stat().st_mtime_ns))
        _rs2.rename_member(_sa / "route_001.pdf", r"C:\durable\route_001.pdf")
        check("rename_member publishes the DURABLE name with the read digest kept",
              any(m.name == r"C:\durable\route_001.pdf" for m in _rs2.members)
              and not any(str(_sa) in m.name for m in _rs2.members))
    finally:
        _rs2.discard()
finally:
    shutil.rmtree(_alias_tmp, ignore_errors=True)

# CMP-AUD-106: a rebuilt comparison that is now CLEAN (no differing columns)
# must not leave prior red evidence surviving at its canonical name.
print("CMP-AUD-106: a clean comparison retires stale prior evidence")
_r6 = Path(tempfile.mkdtemp(prefix="evidence_106_"))
try:
    _cmp6 = _r6 / "hd vs tsn.xlsx"
    _cmp6.write_bytes(b"comparison")
    _wb6, _img6 = ve.sibling_paths(_cmp6)
    _wb6.write_bytes(b"OLD RED evidence workbook")
    _img6.mkdir()
    (_img6 / "old.png").write_bytes(b"old image")
    _note6 = ve._retire_stale_evidence(_wb6, _img6, (), (), None, None)
    check("_retire_stale_evidence removes the stale workbook + image folder",
          not _wb6.exists() and not _img6.exists()
          and "retired" in (_note6 or ""))

    # The 2026-08-05 amendment retired WHOLE LANES, and those cells are
    # refused BEFORE generate() runs — so generate()'s own retirement never
    # fires for them and a pre-amendment set would sit beside its rebuilt
    # comparison looking current. `retire_unsupported` is the sweep the gates
    # call on their way out; without it the orphan survives.
    _cmp6c = _r6 / "excel row vs tsn.xlsx"
    _cmp6c.write_bytes(b"comparison")
    _wb6c, _img6c = ve.sibling_paths(_cmp6c)
    _man6c = ve.evidence_manifest.manifest_path(_cmp6c)
    _man6c.parent.mkdir(parents=True, exist_ok=True)
    _wb6c.write_bytes(b"PRE-AMENDMENT evidence workbook")
    _img6c.mkdir()
    (_img6c / "panel.png").write_bytes(b"a workbook-panel image")
    _man6c.write_text("{}", encoding="utf-8")
    _note6c = ve.retire_unsupported(_cmp6c)
    check("retire_unsupported sweeps a pre-amendment set beside a comparison "
          "whose lane no longer collects evidence",
          not _wb6c.exists() and not _img6c.exists() and not _man6c.exists()
          and "retired" in (_note6c or ""))
    check("...and it is a no-op (empty note) when there is nothing beside the "
          "comparison, so a clean cell is never touched",
          ve.retire_unsupported(_cmp6c) == "" and _cmp6c.exists())
    check("...and it never raises for a comparison that does not exist "
          "(a decoration gate must not fail a comparison)",
          ve.retire_unsupported(_r6 / "no such comparison.xlsx") == "")

    # THE CALL SITES, not just the helper (the mutation lesson: deleting the
    # one line that calls it passed the entire gate). Each retired lane's gate
    # is driven with a planted pre-amendment set and must sweep it.
    class _Ev6c:
        def is_cancelled(self):
            return False

        def on_log(self, _m):
            pass

    def _plant(name):
        cmp_path = _r6 / name
        cmp_path.write_bytes(b"comparison")
        wb, img = ve.sibling_paths(cmp_path)
        wb.write_bytes(b"PRE-AMENDMENT evidence workbook")
        img.mkdir()
        (img / "panel.png").write_bytes(b"a workbook-panel image")
        return cmp_path, wb, img

    _ev_cmp, _ev_wb, _ev_img = _plant("env ramp_summary.xlsx")
    matrix_build._run_env_evidence(
        "ramp_summary", _ev_cmp, {"enabled": True}, _Ev6c(),
        ConsolidateResult(status="ok", summary_lines=[]), None)
    check("the ENV decoration gate sweeps a retired row's pre-amendment set "
          "(ramp_summary — evidence by report type, the third ruling)",
          not _ev_wb.exists() and not _ev_img.exists())

    _sf_cmp, _sf_wb, _sf_img = _plant("self highway_log_pdf.xlsx")
    matrix_build._run_self_evidence(
        "highway_log_pdf", None, None, _sf_cmp, None, {"enabled": True},
        _Ev6c(), ConsolidateResult(status="ok", summary_lines=[]), None)
    check("the SELF decoration gate sweeps its pre-amendment set (the self "
          "lane is retired for every row)",
          not _sf_wb.exists() and not _sf_img.exists())
    # a source-aliased artifact is refused (left in place, noted) not force-removed
    _wb6.write_bytes(b"aliased")
    _note6b = ve._retire_stale_evidence(
        _wb6, _r6 / "absent-images", (_wb6,),
        ve.artifact_store.capture_source_identities((_wb6,)), None, None)
    check("a source-aliased evidence workbook is NOT force-removed",
          _wb6.exists() and "could not remove" in (_note6b or ""))
    _wb6.unlink()

    # generate() on a clean comparison retires the planted prior red evidence.
    # Independent paths from the unit case above so a neutralized retire still
    # reaches these checks (rather than crashing on a lingering folder).
    # CMP-AUD-208: generate() now reads the cells the comparison PUBLISHED, so
    # "clean" has to be a REAL comparison with no counted differences — a
    # stubbed enumerate_diffs can no longer manufacture it. The fixture family
    # is intersection_detail_pdf (the 2026-08-05 ruling: only `_pdf` rows are
    # capable), so the entry gate also wants both print sets present — stub
    # PDFs suffice, the clean path never parses one.
    import compare_intersection_detail_tsn as idt6
    import evidence_intersection_detail as eid6
    import paths as _paths6
    _cmp6b = _r6 / "day2 vs tsn.xlsx"
    _id_row = ["001"] + ["x"] * len(idt6.SHARED_HEADER)
    _id_row[1 + idt6.SHARED_HEADER.index(idt6.KEY)] = "0.100"
    _cons6 = _r6 / "cons.xlsx"; _cons6.write_bytes(b"consolidated")
    _tsn6 = _r6 / "tsn.xlsx"; _tsn6.write_bytes(b"tsn")
    # HF-05: the exact-source binding requires a BOUND comparison — committed
    # generation + typed outcome + provenance over the two side files.
    _checklib.publish_bound_comparison(
        _cmp6b, idt6._SCHEMA, [list(_id_row)], [list(_id_row)],
        (_cons6, _tsn6))
    _wb6b, _img6b = ve.sibling_paths(_cmp6b)
    _tdir6 = _r6 / "tsmis_pdf"; _tdir6.mkdir()
    (_tdir6 / "intersection_detail_route_001.pdf").write_bytes(b"%PDF tsmis")
    _old_lib6 = _paths6.TSN_LIBRARY_ROOT
    _paths6.TSN_LIBRARY_ROOT = _r6 / "tsn_library"
    _libpdf6 = _r6 / "tsn_library" / "intersection_detail" / "pdf"
    _libpdf6.mkdir(parents=True)
    (_libpdf6 / "stub.pdf").write_bytes(b"%PDF tsn")
    _wb6b.write_bytes(b"PRIOR red evidence workbook")
    _img6b.mkdir()
    (_img6b / "prior.png").write_bytes(b"prior image")

    class _Ev106:
        def is_cancelled(self):
            return False

        def on_log(self, _m):
            pass

    _saved106 = (eid6.load_sides, eid6.enumerate_diffs)
    eid6.load_sides = lambda _c, _t: ([], [], {"ok": 1}, None)
    eid6.enumerate_diffs = lambda _a, _b, _s: {}         # a CLEAN comparison
    try:
        _res6 = ve.generate("intersection_detail_pdf", _cons6, _tsn6, _cmp6b,
                            _tdir6, _Ev106())
    finally:
        eid6.load_sides, eid6.enumerate_diffs = _saved106
        _paths6.TSN_LIBRARY_ROOT = _old_lib6
    check("generate() on a clean comparison reports no differing columns",
          "no differing columns" in _res6["note"] and _res6["workbook"] is None)
    check("...and the prior red evidence no longer survives at its canonical name",
          not _wb6b.exists() and not _img6b.exists()
          and "retired" in _res6["note"])
finally:
    shutil.rmtree(_r6, ignore_errors=True)

# RB4-R2-001: a MISSING REQUIRED PRINT is a binding refusal, not a keep-last-
# good skip. The first implementation exited these three cases with a plain
# ValueError BEFORE the retirement path, so the prior workbook, image folder
# AND manifest all survived at their canonical names beside a comparison they
# no longer illustrate. The provenance-stripping fixture above cannot catch it:
# that pair dies inside _bound_provenance, which already retires. This drives
# the shipped generate() with a fully BOUND comparison and withholds one print
# set at a time, so only the print exits can be what fires.
print("RB4-R2-001: a missing print set retires the prior evidence set")


class _SilentEvents:
    def is_cancelled(self):
        return False

    def on_log(self, _m):
        pass


def _print_refusal_case(tsmis_prints, tsn_prints):
    """(survivors, exception) for generate() with the named prints present."""
    root = Path(tempfile.mkdtemp(prefix="evidence_r2_"))
    import compare_intersection_detail_tsn as idt7
    import evidence_intersection_detail as eid7
    import evidence_manifest as em7
    import paths as _paths7
    old_lib = _paths7.TSN_LIBRARY_ROOT
    saved = (eid7.load_sides, eid7.enumerate_diffs)
    try:
        cmp7 = root / "day2 vs tsn.xlsx"
        row7 = ["001"] + ["x"] * len(idt7.SHARED_HEADER)
        row7[1 + idt7.SHARED_HEADER.index(idt7.KEY)] = "0.100"
        cons7 = root / "cons.xlsx"; cons7.write_bytes(b"consolidated")
        tsn7 = root / "tsn.xlsx"; tsn7.write_bytes(b"tsn")
        _checklib.publish_bound_comparison(
            cmp7, idt7._SCHEMA, [list(row7)], [list(row7)], (cons7, tsn7))
        wb7, img7 = ve.sibling_paths(cmp7)
        man7 = em7.manifest_path(cmp7)
        man7.parent.mkdir(parents=True, exist_ok=True)

        tdir7 = root / "tsmis_pdf"; tdir7.mkdir()
        if tsmis_prints:
            (tdir7 / "intersection_detail_route_001.pdf").write_bytes(b"%PDF t")
        _paths7.TSN_LIBRARY_ROOT = root / "tsn_library"
        libpdf7 = root / "tsn_library" / "intersection_detail" / "pdf"
        libpdf7.mkdir(parents=True)
        if tsn_prints:
            (libpdf7 / "stub.pdf").write_bytes(b"%PDF tsn")

        wb7.write_bytes(b"PRIOR evidence workbook")
        img7.mkdir()
        (img7 / "prior.png").write_bytes(b"prior image")
        man7.write_text('{"prior": true}', encoding="utf-8")

        eid7.load_sides = lambda _c, _t: ([], [], {"ok": 1}, None)
        eid7.enumerate_diffs = lambda _a, _b, _s: {}
        raised = None
        try:
            ve.generate("intersection_detail_pdf", cons7, tsn7, cmp7, tdir7,
                        _SilentEvents())
        except Exception as e:                   # noqa: BLE001 — the probe
            raised = e
        return (wb7.exists(), img7.exists(), man7.exists()), raised
    finally:
        eid7.load_sides, eid7.enumerate_diffs = saved
        _paths7.TSN_LIBRARY_ROOT = old_lib
        shutil.rmtree(root, ignore_errors=True)


for _label, _tsmis, _tsn in (("the TSMIS print set", False, True),
                             ("the TSN print set", True, False),
                             ("both print sets", False, False)):
    _survivors, _exc = _print_refusal_case(_tsmis, _tsn)
    check(f"{_label} missing -> no prior workbook, image folder or manifest "
          "survives", not any(_survivors))
    check(f"...and {_label} missing refuses as a BINDING error",
          isinstance(_exc, ve.EvidenceSourceBindingError))

# CMP-AUD-108: a comparison whose ONLY differences live inside a repeated-key
# group must report those columns and say why no image exists — the published
# counts decide, not the adapter's (correctly) empty candidate list. Driven
# through the shipped generate(), and it must not touch a PDF to find out.
print("CMP-AUD-108: duplicate-only differences are reported, never zeroed")
_r8 = Path(tempfile.mkdtemp(prefix="evidence_108_"))
try:
    import compare_intersection_detail_tsn as idt8
    import evidence_intersection_detail as eid8
    import paths as _paths8
    _cmp8 = _r8 / "dup only vs tsn.xlsx"
    _base8 = ["001"] + ["x"] * len(idt8.SHARED_HEADER)
    _base8[1 + idt8.SHARED_HEADER.index(idt8.KEY)] = "0.100"
    _desc8 = 1 + idt8.SHARED_HEADER.index("Description")

    def _dup_row(desc):
        row = list(_base8)
        row[_desc8] = desc
        return row

    _cons8 = _r8 / "cons.xlsx"; _cons8.write_bytes(b"consolidated")
    _tsn8 = _r8 / "tsn.xlsx"; _tsn8.write_bytes(b"tsn")
    _checklib.publish_bound_comparison(
        _cmp8, idt8._SCHEMA,
        [_dup_row("A1"), _dup_row("A2")], [_dup_row("B1"), _dup_row("B2")],
        (_cons8, _tsn8))
    _tdir8 = _r8 / "tsmis_pdf"; _tdir8.mkdir()
    (_tdir8 / "intersection_detail_route_001.pdf").write_bytes(b"%PDF tsmis")
    _old_lib8 = _paths8.TSN_LIBRARY_ROOT
    _paths8.TSN_LIBRARY_ROOT = _r8 / "tsn_library"
    _libpdf8 = _r8 / "tsn_library" / "intersection_detail" / "pdf"
    _libpdf8.mkdir(parents=True)
    (_libpdf8 / "stub.pdf").write_bytes(b"%PDF tsn")
    _addressed8 = []
    _saved8 = (eid8.load_sides, eid8.enumerate_diffs,
               ve._locate_tsmis_sources)
    eid8.load_sides = lambda _c, _t: ([], [], {"ok": 1}, None)
    # The real adapter drops duplicate keys, so it proposes nothing here.
    eid8.enumerate_diffs = lambda _a, _b, _s: {}
    ve._locate_tsmis_sources = (
        lambda *a, **k: _addressed8.append(1) or ({}, {}, set()))
    try:
        _res8 = ve.generate("intersection_detail_pdf", _cons8, _tsn8, _cmp8,
                            _tdir8, _Ev106())
    finally:
        (eid8.load_sides, eid8.enumerate_diffs,
         ve._locate_tsmis_sources) = _saved8
        _paths8.TSN_LIBRARY_ROOT = _old_lib8
    check("the published difference count is reported, not zero",
          _res8["fields_with_diffs"] == 1 and "2 published difference"
          in _res8["note"])
    check("...and it never claims there are no differing columns",
          "no differing columns" not in _res8["note"])
    check("the unrenderable column is NAMED with its reason",
          "repeated-key groups" in (_res8["misses"].get("Description") or ""))
    check("no source was addressed to discover that", not _addressed8)
    check("the ledger comes back with the result",
          _res8["ledger"].difference_cells == 2
          and _res8["ledger"].duplicate_groups == 1
          and len(_res8["ledger_digest"]) == 64)
finally:
    shutil.rmtree(_r8, ignore_errors=True)

# CMP-AUD-109: the workbook + images publish as ONE set. When the workbook can't
# reach its canonical name (locked open in Excel), the images are diverted too,
# so the canonical pair is never a NEW workbook beside OLD images.
print("CMP-AUD-109: workbook + images publish as one set")
_r9 = Path(tempfile.mkdtemp(prefix="evidence_109_"))
try:
    from PIL import Image as _PILImage9

    def _publish_case(name, lock_workbook):
        cmp9 = _r9 / f"{name}.xlsx"
        cmp9.write_bytes(b"comparison")
        wb9, img9 = ve.sibling_paths(cmp9)
        wb9.write_bytes(b"OLD workbook")
        img9.mkdir()
        (img9 / "old.png").write_bytes(b"old image")
        tmp9 = _r9 / f"{name}.rendered"
        tmp9.mkdir()
        _PILImage9.new("RGB", (30, 15), "white").save(tmp9 / "Description_1_pair.png")
        entries = [{"field": "Description", "route": "001", "key": "1.0",
                    "va": "A", "vb": "B", "note": "",
                    "pair": "Description_1_pair.png"}]
        info = {"report": "HD", "comparison": cmp9.name, "examples": 1,
                "seed": "00000000", "tsmis_dir": "A", "tsn_dir": "B"}
        captured = ve.artifact_store.capture_source_identities((cmp9,))
        real_replace = ve.os.replace
        state = {"n": 0}

        def replace(src, dst):
            if lock_workbook and Path(dst) == wb9 and state["n"] == 0:
                state["n"] += 1
                raise PermissionError("simulated workbook open in Excel")
            return real_replace(src, dst)

        ve.os.replace = replace
        try:
            res = ve._publish_evidence_set(
                wb9, img9, tmp9, entries, {}, info, "pair", (cmp9,),
                captured, None, None, ve.owned_dir.directory_identity(tmp9))
        finally:
            ve.os.replace = real_replace
        return wb9, img9, res

    _wb9, _img9, _res9 = _publish_case("happy", lock_workbook=False)
    check("no lock -> both promoted to canonical",
          _res9["status"] == "promoted"
          and _res9["workbook"] == str(_wb9) and _res9["folder"] == str(_img9)
          and _wb9.read_bytes() != b"OLD workbook"          # workbook replaced
          and not (_img9 / "old.png").exists())             # images replaced

    _wb9b, _img9b, _res9b = _publish_case("locked", lock_workbook=True)
    check("workbook locked -> the SET diverts, status honest (not success)",
          _res9b["status"] == "diverted"
          and _res9b["workbook"] is None and _res9b["folder"] is None)
    check("...the OLD workbook stays at canonical (not overwritten)",
          _wb9b.read_bytes() == b"OLD workbook")
    check("...and the OLD images STAY at canonical (never a new-wb / old-images mix)",
          (_img9b / "old.png").exists()
          and (_img9b / "old.png").read_bytes() == b"old image")
    check("...the new set lands in .new siblings",
          any(p.name.startswith(_wb9b.stem + ".new-")
              for p in _r9.glob(_wb9b.stem + ".new-*.xlsx"))
          and any(p.is_dir() for p in _r9.glob(_img9b.name + ".new-*")))
finally:
    shutil.rmtree(_r9, ignore_errors=True)

# The strip crop is a FULL-WIDTH page band stretched over the cell box
# (v0.26.0): the adapters' xspan covers only the record's own words, so a crop
# keyed to it clipped a blank cell's red box (drawn where the value WOULD
# print) and truncated the neighbors' longer text — the HSL clipped-box defect.
_cw = ve._crop_window(2000, 3000, (500, 100, 700, 110), (100, 110))
check("crop is full page width, record band ± the vertical context",
      _cw == (0, int((100 - ve._CTX_PT) * ve._SC), 2000,
              int((110 + ve._CTX_PT + 2) * ve._SC)))
_cw2 = ve._crop_window(2000, 3000, (500, 60, 700, 180), (100, 110))
check("a cell box taller than the record stretches the band over it",
      _cw2[1] == int(56 * ve._SC) and _cw2[3] == int(184 * ve._SC))
check("the band clamps to the page edges",
      ve._crop_window(2000, 300, (0, 0, 10, 10), (0, 4))[1] == 0
      and ve._crop_window(2000, 300, (0, 100, 10, 110), (110, 118))[3] == 300)
# The quote-characters clarifier: a diff whose values differ ONLY in quote
# characters ('' vs " vs ') prints near-identically, so the evidence header
# must SAY the difference is real and name both sides' characters (the
# censused case: Intersection Detail KER 046 @ 50.904).
_qn = ve._quote_note("''F'' ST", '"F" ST')
check("quote-only diff -> note names both sides' quote characters",
      "quote characters only" in _qn and "two apostrophes" in _qn
      and "a quotation mark" in _qn)
check("quote-only detection is direction- and kind-aware",
      "one apostrophe" in ve._quote_note("'F' ST", '"F" ST')
      and "TSMIS prints \" (a quotation mark)" in ve._quote_note('"F" ST', "''F'' ST"))
check("genuinely different values -> no note",
      ve._quote_note("MYRTLE ST", "SANFORD AVE") == ""
      and ve._quote_note("''F'' ST", '"F" AVE') == "")
check("identical values (incl. both blank) -> no note",
      ve._quote_note("''E'' ST", "''E'' ST") == ""
      and ve._quote_note(None, "") == "")
avail = ve.availability()
check("availability shape (rows/tsn_pdfs/ready/dir/reports/row_reports/deps_ok)",
      set(avail) >= {"rows", "tsn_pdfs", "ready", "dir", "reports", "row_reports",
                     "deps_ok"})
check("availability reports the five supported evidence families, per-dir + "
      "source kind (HD joined in v0.37.0 — its statewide-XLSX TSN library keeps "
      "the separate optional pdf/ drop, like Intersection/Ramp Detail)",
      [r["key"] for r in avail["reports"]]
      == ["highway_detail", "highway_log", "highway_sequence",
          "intersection_detail", "ramp_detail"]
      and all(set(r) >= {"key", "label", "tsn_pdfs", "dir", "source"}
              for r in avail["reports"])
      and {r["key"]: r["source"] for r in avail["reports"]}
      == {"highway_detail": "pdf", "highway_log": "raw",
          "highway_sequence": "raw",
          "intersection_detail": "pdf", "ramp_detail": "pdf"})
check("row_reports maps every capable row to its report (the per-cell action's gate)",
      avail["row_reports"] == ve.TSN_PDF_REPORT
      and set(avail["row_reports"]) == set(ve.rows()))
# The env cells carry their own per-route prints, so readiness stays the
# imaging deps alone (a vs-TSN cell missing its TSN prints reports that per
# cell); env_rows is the LITERAL four `_pdf` placements — Ramp Summary is
# REMOVED by the third ruling (report-type rule).
check("ready == deps alone and env_rows is the LITERAL four `_pdf` "
      "placements (pinned, not self-derived — RB-4 audit; RS removed)",
      avail["ready"] == avail["deps_ok"]
      and avail.get("env_rows") == ["highway_log_pdf",
                                    "highway_sequence_pdf",
                                    "intersection_detail_pdf",
                                    "ramp_detail_pdf"])
check("ramp_summary has NO evidence lane left (env_capable False too)",
      not ve.env_capable("ramp_summary") and not ve.capable("ramp_summary"))

print("caller-side gate (matrix_build.evidence_opts_for)")
check("toggle off -> None",
      matrix_build.evidence_opts_for(None, "highway_log_pdf", lambda s: s) is None
      and matrix_build.evidence_opts_for({"enabled": False, "examples": 5},
                                         "highway_log_pdf", lambda s: s) is None)
check("unsupported row -> None (Excel rows and ramp_summary alike)",
      matrix_build.evidence_opts_for({"enabled": True}, "ramp_summary",
                                     lambda s: s) is None
      and matrix_build.evidence_opts_for({"enabled": True}, "highway_log",
                                         lambda s: s) is None
      and matrix_build.evidence_opts_for({"enabled": True}, "highway_detail",
                                         lambda s: s) is None)
opts = matrix_build.evidence_opts_for({"enabled": True, "examples": 99},
                                      "highway_log_pdf",
                                      lambda s: Path("cell") / s)
check("supported row -> resolved PDF dir + clamped examples + default layout",
      opts == {"tsmis_pdf_dir": Path("cell") / "highway_log_pdf",
               "examples": 10, "layout": "pair"})
check("evidence_opts_for carries the chosen layout (unknown -> the 'pair' default)",
      matrix_build.evidence_opts_for(
          {"enabled": True, "layout": "both"}, "highway_log_pdf",
          lambda s: s)["layout"] == "both"
      and matrix_build.evidence_opts_for(
          {"enabled": True, "layout": "nonsense"}, "highway_log_pdf",
          lambda s: s)["layout"] == "pair")

# --------------------------------------------------------------------------- #
print("layout choice + per-column tabs (the workbook structure)")
check("LAYOUTS + default mirror settings.get_evidence_layout",
      ve.LAYOUTS == ("pair", "stacked", "both") and ve.DEFAULT_LAYOUT == "pair")
check("normalize_layout: known pass-through, unknown/None -> pair",
      (ve.normalize_layout("pair"), ve.normalize_layout("stacked"),
       ve.normalize_layout("both"), ve.normalize_layout(None),
       ve.normalize_layout("x")) == ("pair", "stacked", "both", "pair", "pair"))
check("_layout_keys: pair->(pair,), stacked->(stacked,), both->(stacked,pair)",
      ve._layout_keys("pair") == ("pair",)
      and ve._layout_keys("stacked") == ("stacked",)
      and ve._layout_keys("both") == ("stacked", "pair"))
_used = {"Summary"}
_n1 = ve._sheet_name("Med V/WDA", _used)            # '/' is illegal in a name
_n2 = ve._sheet_name("Med V/WDA", _used)            # collision -> disambiguated
_n3 = ve._sheet_name("Distance To Next Point", _used, " (side-by-side)")
check("sheet names are legal (no []:*?/\\), <=31, unique, suffix-reserved",
      not (set(_n1) & set("[]:*?/\\")) and _n1 != _n2
      and len(_n3) <= 31 and _n3.endswith(" (side-by-side)")
      and _n1 in _used and _n2 in _used)

_wbtmp = Path(tempfile.mkdtemp(prefix="evidence_layout_"))
try:
    from PIL import Image as _PILImage
    _imgs = _wbtmp / "imgs"
    _imgs.mkdir()

    def _png(name):
        _PILImage.new("RGB", (40, 20), (255, 255, 255)).save(_imgs / name)
        return name

    def _entry(field, route, va, vb):
        return {"field": field, "route": route, "key": f"{route}.0",
                "va": va, "vb": vb, "note": "",
                "stacked": _png(f"{field}_{route}_s.png"),
                "pair": _png(f"{field}_{route}_p.png")}

    _entries = [_entry("Description", "001", "A", "B"),
                _entry("Description", "002", "C", "D"),
                _entry("FT", "003", "1", "2")]
    _info = {"report": "Probe", "comparison": "c.xlsx", "examples": 2,
             "seed": "00000000", "tsmis_dir": "A", "tsn_dir": "B"}

    _wb_pair = _wbtmp / "p (evidence).xlsx"
    ve._write_workbook(_wb_pair, _imgs, _entries,
                       {"HG": "no verifiable example — ambiguous"}, _info,
                       layout="pair")
    _pw = load_workbook(_wb_pair)
    check("pair layout -> Summary + ONE tab per differing column, no suffix",
          _pw.sheetnames == ["Summary", "Description", "FT"])
    check("Summary lists only the rendered (side-by-side) image per row",
          _pw["Summary"]["E7"].value and ".png" in _pw["Summary"]["E7"].value
          and "  /  " not in _pw["Summary"]["E7"].value)
    _pw.close()

    _wb_stacked = _wbtmp / "s (evidence).xlsx"
    ve._write_workbook(_wb_stacked, _imgs, _entries, {}, _info, layout="stacked")
    check("stacked layout -> per-column tabs, no suffix",
          load_workbook(_wb_stacked).sheetnames == ["Summary", "Description", "FT"])

    _wb_both = _wbtmp / "b (evidence).xlsx"
    ve._write_workbook(_wb_both, _imgs, _entries, {}, _info, layout="both")
    check("both layout -> a stacked AND a side-by-side tab per column",
          load_workbook(_wb_both).sheetnames
          == ["Summary", "Description (stacked)", "FT (stacked)",
              "Description (side-by-side)", "FT (side-by-side)"])

    # A pair-only entry (as generate() produces for layout='pair') has no
    # 'stacked' key; a stacked tab request must simply skip it.
    _pair_only = [{"field": "HG", "route": "009", "key": "9.0", "va": "X",
                   "vb": "Y", "note": "", "pair": _png("HG_009_p.png")}]
    _wb_po = _wbtmp / "po (evidence).xlsx"
    ve._write_workbook(_wb_po, _imgs, _pair_only, {}, _info, layout="stacked")
    check("a layout with no rendered image for a column makes no empty tab",
          load_workbook(_wb_po).sheetnames == ["Summary"])
finally:
    shutil.rmtree(_wbtmp, ignore_errors=True)

# --------------------------------------------------------------------------- #
print("adapter LOCKSTEP pins vs the PDF consolidator")
check("window shapes: 10-cell line 1 + 25-cell line 2 == the 34 columns",
      chd.N_COLS_L1 == 10 and chd.N_COLS_L2 == 25 and len(hdc.HEADER) == 34
      and 9 + chd.N_COLS_L2 == len(hdc.HEADER))
check("postmile token regex accepts the glued forms the walk classifies on",
      all(chd.PM_TOKEN_RE.match(t)
          for t in ("S000.000", "000.000E", "R012.243R", "C043.925R"))
      and not chd.PM_TOKEN_RE.match("11 IMP 007"))
check("date-token guard: TASAS date yes, page-header date no",
      bool(chd.DATE_TOKEN_RE.search("64-01-01"))
      and not chd.DATE_TOKEN_RE.search("2026-07-07"))
check("FIELDS = every shared column except the key (PS included)",
      ehd.FIELDS == [f for f in cht.SHARED_HEADER if f != "Post Mile"]
      and "PS" in ehd.FIELDS and len(ehd.FIELDS) == 34)
check("TSN group map covers exactly FIELDS",
      set(ehd.TSN_GROUP) == set(ehd.FIELDS))
check("RB half of the TSN map is MIRRORED (inner before width before outer)",
      (ehd.TSN_GROUP["RB IN-TO"], ehd.TSN_GROUP["RB Wid"],
       ehd.TSN_GROUP["RB OT-TO"]) == ("rbto1", "rbwid", "rbto2"))

print("TSN print regexes on realistic two-line records")
l1 = "R 004.972E  000.123  11-08-01  D  F  Y15-05-18  LGNB  R  22-01-01  054062  S"
m1 = ehd.L1_RE.match(l1)
check("line 1: prefix + equation marker + sig-flagged eff + city all parse",
      bool(m1) and (m1.group("pp").strip(), m1.group("mile"), m1.group("ps"),
                    m1.group("city"), m1.group("ru"), m1.group("beg"))
      == ("R", "004.972", "E", "LGNB", "R", "22-01-01"))
l1b = "000.000  000.055  64-01-01  U  C  64-01-01  B  21-01-01  242400"
m1b = ehd.L1_RE.match(l1b)
check("line 1: bare PM, no city, no marker",
      bool(m1b) and m1b.group("pp") is None and m1b.group("ps") is None
      and m1b.group("city") is None and m1b.group("ru") == "B")
l2 = ("SANDHILLS DITCH  A  Y90-03-15  C  5  N  8  8  64  8  8  "
      "*90-03-15  H  7  F  12V  Y85-12-27  C  4  N  2  2  44  8  8")
m2 = ehd.L2_RE.match(l2)
check("line 2: desc + NA + the three sig-flagged blocks parse",
      bool(m2) and (m2.group("desc"), m2.group("na"), m2.group("lbeff"),
                    m2.group("medwda"), m2.group("rbto2"))
      == ("SANDHILLS DITCH", "A", "90-03-15", "12V", "8"))
l2e = ("A  Y90-03-15  C  5  N  8  8  64  8  8  "
       "*90-03-15  H  7  F  12V  Y85-12-27  C  4  N  2  2  44  8  8")
m2e = ehd.L2_RE.match(l2e)
check("line 2: EMPTY description still parses (the \\s* fix)",
      bool(m2e) and m2e.group("desc") == "" and m2e.group("na") == "A")
# the REAL fully-dittoed right-roadbed block from the D04 print (route 237 @
# R008.816L): width-matched '+' runs — an 8-char run for the dittoed eff DATE,
# '+++' for the 3-digit width.
l2d = ("EB 37-84K A 02-12-09 H 3 N 10 10 36 10 10 "
       "02-12-09 H 7 E 30V ++++++++ + ++ + ++ ++ +++ ++ ++")
m2d = ehd.L2_RE.match(l2d)
check("line 2: TSN width-matched DITTO runs parse (dates included)",
      bool(m2d) and (m2d.group("rbeff"), m2d.group("rbt"), m2d.group("rbln"),
                     m2d.group("rbwid"), m2d.group("rbtr2"))
      == ("++++++++", "+", "++", "+++", "++")
      and ehd.project("RB #Ln", m2d.group("rbln")) == "++"
      and ehd.project("RB Eff", m2d.group("rbeff")) == "++++++++")

print("span→box math (word-indexed line)")
ln = {"text": "AA BBB C", "offs": [(0, 2, {"x0": 10.0, "x1": 20.0}),
                                   (3, 6, {"x0": 30.0, "x1": 45.0}),
                                   (7, 8, {"x0": 55.0, "x1": 60.0})]}
check("value span boxes its words", ehd._span_box(ln, 3, 6) == (30.0, 45.0))
check("empty span boxes the neighbor gap",
      ehd._span_box(ln, 3, 3) == (21.0, 29.0))

print("verification projections")
check("PS is marker-derived", ehd.project("PS", "E") == "E"
      and ehd.project("PS", "") == "" and ehd.project("PS", None) == "")
check("other fields ride the comparator's own projection (NA fold, WDA glue)",
      ehd.project("NA", "A") == "" and ehd.project("NA", "N") == "N"
      and ehd.project("Med V/WDA", "8V") == "08V")

# --------------------------------------------------------------------------- #
print("diff enumeration (unique keys + sidecar)")
def _row(route, key, **over):
    r = [route] + [""] * len(cht.SHARED_HEADER)
    r[1 + cht.SHARED_HEADER.index("Post Mile")] = key
    for f, v in over.items():
        r[1 + cht.SHARED_HEADER.index(f)] = v
    return r

a_rows = [_row("001", "001.000", **{"LB Wid": "24"}),
          _row("001", "002.000", **{"LB Wid": "24"}),   # dup key: excluded
          _row("001", "002.000", **{"LB Wid": "25"}),
          _row("001", "003.000", **{"AC": "F"})]
b_rows = [_row("001", "001.000", **{"LB Wid": "26"}),
          _row("001", "002.000", **{"LB Wid": "24"}),
          _row("001", "003.000", **{"AC": "F"})]
sc = {("001", "001.000"): [("06", "TUL")], ("001", "003.000"): [("06", "TUL")]}
diffs = ehd.enumerate_diffs(a_rows, b_rows, sc)
check("only the unique-key LB Wid diff is enumerated, with its district",
      list(diffs) == ["LB Wid"] and len(diffs["LB Wid"]) == 1
      and diffs["LB Wid"][0]["key"] == "001.000"
      and (diffs["LB Wid"][0]["dist"], diffs["LB Wid"][0]["cnty"]) == ("06", "TUL")
      and (diffs["LB Wid"][0]["va"], diffs["LB Wid"][0]["vb"]) == ("24", "26"))

# CMP-AUD-107: HD evidence now judges each cell with the comparison's OWN
# compared_cell verdict (Excel TRIM + the Med V/WDA fold), so a non-difference
# the workbook does NOT count can never be enumerated. Raw string compare
# invented both of these before.
_ws_a = [_row("002", "010.000", **{"HG": "A  B", "Med V/WDA": "06V", "AC": "F"})]
_ws_b = [_row("002", "010.000", **{"HG": "A B", "Med V/WDA": "6V", "AC": "G"})]
_ws_diffs = ehd.enumerate_diffs(_ws_a, _ws_b, {("002", "010.000"): [("06", "TUL")]})
check("HG internal-whitespace-only difference is NOT enumerated (Excel TRIM)",
      "HG" not in _ws_diffs)
check("Med V/WDA leading-zero-only difference is NOT enumerated (Med-Wid fold)",
      "Med V/WDA" not in _ws_diffs
      and cht._SCHEMA.is_medwid(cht.SHARED_HEADER.index("Med V/WDA")))
check("a genuinely different asserting cell (AC) IS still enumerated, TRIM display",
      list(_ws_diffs) == ["AC"]
      and (_ws_diffs["AC"][0]["va"], _ws_diffs["AC"][0]["vb"]) == ("F", "G"))
# the projection is TRIM-aligned so a real diff still round-trips against the
# compared_cell display (the verification consistency half of CMP-AUD-107)
check("HD project() applies the Excel TRIM so a verified value matches the display",
      ehd.project("Description", "A  B") == "A B"
      and ehd.project("Med V/WDA", "8V") == "08V" and ehd.project("PS", "E") == "E")

# --------------------------------------------------------------------------- #
print("TSN loader sidecar contract")
tmp = Path(tempfile.mkdtemp())
try:
    raw_dir = tmp / "raw"
    raw_dir.mkdir()
    raw = raw_dir / "raw.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = cht.TSN_SHEET
    cols = list(cht.TSN_RAW_HEADER)
    ws.append(cols)
    base = {c: "" for c in cols}
    base.update(DIST="06", CNTY="TUL.", RTE="99", PP="R", POSTMILE="004.972",
                E_IND="E", HG="R", LENGTH="0.123", NON_ADD="A", M_WID="8",
                M_VA="V", DESCRIPTION="X  Y")
    ws.append([base[c] for c in cols])
    wb.save(raw)
    wb.close()

    rows_locked = cht.tsn_rows_from_raw(raw)
    rows_dcr, dcr = tlh.tsn_rows_with_dcr(raw)
    check("tsn_rows_with_dcr rows are IDENTICAL to the locked loader's",
          rows_dcr == rows_locked and len(rows_dcr) == 1)
    check("…and the sidecar carries (district, county-dot-stripped)",
          dcr == [("06", "TUL")])

    # the normalized library sheet: shared header + EXACTLY the sidecar columns
    out = tmp / "norm.xlsx"
    res = tlh.build_into(raw_dir, out, events=None, confirm_overwrite=lambda p: True)
    nwb = load_workbook(out)
    nws = nwb[cht.NORMALIZED_SHEET]
    hdr = [c.value for c in nws[1]]
    first = [c.value for c in nws[2]]
    nwb.close()
    check("normalized header = Route + shared + sidecar",
          res.status == "ok"
          and hdr == ["Route"] + cht.SHARED_HEADER + tlh.SIDECAR_HEADER)
    check("normalized row carries the sidecar values at the tail",
          first[-2:] == ["06", "TUL"] and first[0] == "099")

    # load_sides reads the sidecar back; the comparator side stays shared-width
    a_cons = tmp / "cons.xlsx"
    cw = Workbook()
    cs = cw.active
    cs.title = cht.TSMIS_SHEET
    cs.append(list(cht._TSMIS_HEADER))       # CMP-AUD-034 exact consolidated header
    cs.append(["099", "R004.972R", "000.123"] + [""] * 32)
    cw.save(a_cons)
    cw.close()
    ar, br, sc2, note = ehd.load_sides(a_cons, out)
    check("load_sides: rows in comparator shape, sidecar keyed by (route,key)",
          note is None and len(ar) == 1 and len(br) == 1
          and len(br[0]) == 1 + len(cht.SHARED_HEADER)
          and sc2.get(("099", br[0][1])) == [("06", "TUL")])
    check("both sides land on the same canonical key (roadbed-aware)",
          ar[0][1] == br[0][1] == "R004.972R")

    # a LEGACY normalized library (no sidecar) is refused with the rebuild hint
    old = tmp / "old.xlsx"
    ow = Workbook()
    os_ = ow.active
    os_.title = cht.NORMALIZED_SHEET
    os_.append(["Route"] + cht.SHARED_HEADER)
    os_.append(br[0])
    ow.save(old)
    ow.close()
    _a, _b, sc3, note3 = ehd.load_sides(a_cons, old)
    check("legacy library -> sidecar None + 'rebuild the TSN library' hint",
          sc3 is None and note3 and "rebuild the TSN library" in note3)
finally:
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)

# --------------------------------------------------------------------------- #
print("Intersection Detail adapter (v0.22.0): maps + windows + LOCKSTEP")
import compare_intersection_detail_tsn as idt                # noqa: E402
import consolidate_tsmis_intersection_detail_pdf as idpdf    # noqa: E402
import evidence_intersection_detail as eid                   # noqa: E402
import tsn_load_intersection_detail as tli                   # noqa: E402

check("ID FIELDS = every shared column except the key (34 — District/County "
      "joined per ID-79; Route Suffix included)",
      eid.FIELDS == [f for f in idt.SHARED_HEADER if f != idt.KEY]
      and "Route Suffix" in eid.FIELDS and "District" in eid.FIELDS
      and "County" in eid.FIELDS and len(eid.FIELDS) == 34)
check("ID TSMIS cell map covers exactly FIELDS",
      set(eid._TSMIS_CELL) == set(eid.FIELDS))
check("ID TSN cell map covers exactly FIELDS",
      set(eid.TSN_CELL) == set(eid.FIELDS))
_l1n = {n for n, _lo, _hi in eid._L1_WIN}
_l2n = {n for n, _lo, _hi in eid._L2_WIN}
check("every TSN cell target has a fixed window on its line",
      all((n in _l1n if ln == 1 else n in _l2n)
          for ln, n in eid.TSN_CELL.values()))
check("ID TSMIS value positions mirror the comparator's (consolidated - Route)",
      eid._TSMIS_SRC == {f: p - 1 for f, p in idt._TSMIS_POS.items()})
check("Xing Line Lgth: TSMIS boxes rowB window 17, TSN boxes LINE 1's X-OVR "
      "(each side its own print position)",
      eid._TSMIS_CELL["Xing Line Lgth"] == (2, 17)
      and eid.TSN_CELL["Xing Line Lgth"] == (1, "X_CROSS_OVERRIDE"))
check("the Intrte swap mirrored: TSMIS boxes Route at rowB window 12",
      eid._TSMIS_CELL["Intrte Route"] == (2, 12)
      and eid._TSMIS_CELL["Intrte PM Suffix"] == (2, 16))
check("Route Suffix boxes the Location cell on both sides",
      eid._TSMIS_CELL["Route Suffix"] == (1, 3)
      and eid.TSN_CELL["Route Suffix"] == (1, "LOC"))
check("LOCKSTEP handles the consolidator's own pieces (rowA/rowB discriminators)",
      idpdf._is_rowA(["", "000.204", "", "12 ORA 001"] + [""] * 17)
      and bool(idpdf.INT_ROWB_RE.match("11050"))
      and bool(idpdf.OLD_PM_RE.match("0.204")))

print("ID TSN print: fixed windows, max-overlap, flag strip, LOC tokens")
check("LOC tokenizer: 3-char / dotted / 2-char counties + a route suffix",
      bool(eid._LOC_RE.match("12 ORA 001")) and bool(eid._LOC_RE.match("04 CC. 004"))
      and bool(eid._LOC_RE.match("07 LA 001")) and bool(eid._LOC_RE.match("07 LA 210U"))
      and not eid._LOC_RE.match("NB ON FROM SB RTE 5"))
_w1 = [{"t": "R", "x0": 14.0, "x1": 19.0}, {"t": "000.204", "x0": 25.0, "x1": 59.0},
       {"t": "12", "x0": 72.0, "x1": 82.0}, {"t": "ORA", "x0": 86.0, "x1": 101.0},
       {"t": "210U", "x0": 106.0, "x1": 125.0},
       {"t": "Y91-08-24", "x0": 406.0, "x1": 454.0}]
_a1 = eid._assign_win(_w1, eid._L1_WIN)
check("LOCATION is ONE window (a 2-char county can't shift the route out of it)",
      _a1["LOC"][0] == "12 ORA 210U")
check("max-overlap: a signature-flagged date leaning left stays in its DATE window",
      _a1["EFF_DATE_LT"][0] == "Y91-08-24" and _a1["TY_CT"][0] == "")
_l1 = {"page": 3, "words": _w1, "top": 100.0, "bottom": 110.0}
_w2 = [{"t": "JCT", "x0": 72.0, "x1": 86.0}, {"t": "5", "x0": 90.0, "x1": 95.0}]
_l2 = {"page": 3, "words": _w2, "top": 111.0, "bottom": 121.0}
_rec = {"l1": _l1, "a1": _a1, "l2": _l2, "a2": eid._assign_win(_w2, eid._L2_WIN),
        "dist": "12"}
check("the glued flag is stripped from the VALUE ('Y91-08-24' -> 1991-08-24)…",
      eid._tsn_raw(_rec, "Lighting Eff-Date") == "91-08-24"
      and eid.tsn_value(_rec, "Lighting Eff-Date") == "1991-08-24")
_pg, _box, _yspan, _xspan = eid.tsn_box(_rec, "Lighting Eff-Date")
check("…while the BOX keeps the printed token (flag included)",
      _pg == 3 and _box[0] <= 406.0 and _box[2] >= 454.0)
_pg2, _box2, _y2, _x2 = eid.tsn_box(_rec, "Int St Eff-Date")
check("a BLANK cell boxes its fixed template window (the window IS the cell)",
      _pg2 == 3 and 405 <= _box2[0] <= 415 and 440 <= _box2[2] <= 460
      and _box2[1] < _box2[3])
check("Route Suffix reads the LOC route token ('210U' -> 'U')",
      eid.tsn_value(_rec, "Route Suffix") == "U")

print("ID diff enumeration: unique keys, sidecar, the comparison's own trim")
def _idrow(route, pm, **over):
    r = [route] + [""] * len(idt.SHARED_HEADER)
    r[1 + idt.KEY_FIELD] = pm
    for f, v in over.items():
        r[1 + idt.SHARED_HEADER.index(f)] = v
    return r

_ar = [_idrow("001", "0.204", HG="D", Description="A  B"),
       _idrow("001", "1.000", HG="D"),      # dup key: excluded
       _idrow("001", "1.000", HG="U")]
_br = [_idrow("001", "0.204", HG="U", Description="A B"),
       _idrow("001", "1.000", HG="D")]
_sc = {("001", "0.204"): [("12", "ORA")]}
_diffs = eid.enumerate_diffs(_ar, _br, _sc)
check("only the unique-key HG diff is enumerated, with its district/county",
      list(_diffs) == ["HG"] and len(_diffs["HG"]) == 1
      and (_diffs["HG"][0]["dist"], _diffs["HG"][0]["cnty"]) == ("12", "ORA")
      and (_diffs["HG"][0]["va"], _diffs["HG"][0]["vb"]) == ("D", "U"))
check("a whitespace-run-only difference is NOT enumerated (compare_core's trim)",
      "Description" not in _diffs)

print("ID TSN loader sidecar contract")
tmp2 = Path(tempfile.mkdtemp())
try:
    raw_dir2 = tmp2 / "raw"
    raw_dir2.mkdir()
    raw2 = raw_dir2 / "raw.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = idt.TSN_SHEET
    cols2 = list(idt.TSN_RAW_HEADER)
    ws2.append(cols2)
    _b2 = {c: "" for c in cols2}
    _b2.update(PP="R", POST_MILE=" 000.204", LOCATION="04 CC. 004",
               DATE_REC="73-10-19", HG="D", RU="U", X_CROSS_OVERRIDE="0250")
    ws2.append([_b2[c] for c in cols2])
    wb2.save(raw2)
    wb2.close()

    rows_locked2 = idt.tsn_rows_from_raw(raw2)
    rows_dcr2, dcr2 = tli.tsn_rows_with_dcr(raw2)
    check("tsn_rows_with_dcr rows are IDENTICAL to the locked loader's",
          rows_dcr2 == rows_locked2 and len(rows_dcr2) == 1)
    check("…and the sidecar carries (district, county-dot-stripped)",
          dcr2 == [("04", "CC")])

    out2 = tmp2 / "norm.xlsx"
    res2 = tli.build_into(raw_dir2, out2, events=None, confirm_overwrite=lambda p: True)
    nwb2 = load_workbook(out2)
    nws2 = nwb2[idt.NORMALIZED_SHEET]
    hdr2 = [c.value for c in nws2[1]]
    first2 = [c.value for c in nws2[2]]
    nwb2.close()
    check("normalized header = Route + shared + sidecar (v3 shape, XLL included)",
          res2.status == "ok"
          and hdr2 == ["Route"] + idt.SHARED_HEADER + tli.SIDECAR_HEADER
          and "Xing Line Lgth" in hdr2 and "ML 2nd Eff-Date" not in hdr2)
    check("normalized row carries the sidecar values at the tail",
          first2[-2:] == ["04", "CC"] and first2[0] == "004")

    a_cons2 = tmp2 / "cons.xlsx"
    cw2 = Workbook()
    cs2 = cw2.active
    cs2.title = idt.TSMIS_SHEET
    cs2.append(list(idt._TSMIS_HEADER))      # CMP-AUD-034 exact consolidated header
    _r2 = [None] * 36
    _r2[0], _r2[1], _r2[2], _r2[4] = "004", "R", "000.204", "04 CC. 004"
    cs2.append(_r2)
    cw2.save(a_cons2)
    cw2.close()
    ar2, br2, sc22, note2 = eid.load_sides(a_cons2, out2)
    check("load_sides: rows in comparator shape, sidecar keyed by (route,key)",
          note2 is None and len(ar2) == 1 and len(br2) == 1
          and len(br2[0]) == 1 + len(idt.SHARED_HEADER)
          and sc22.get(("004", br2[0][1 + idt.KEY_FIELD])) == [("04", "CC")])
    # CMP-AUD-045: the keys are county-aware PhysicalKeys — identity-equal
    # across sides, displaying the normalized PM text.
    check("both sides land on the same physical key (display '0.204')",
          ar2[0][1 + idt.KEY_FIELD] == br2[0][1 + idt.KEY_FIELD]
          and str(ar2[0][1 + idt.KEY_FIELD]) == "0.204")

    old2 = tmp2 / "old.xlsx"
    ow2 = Workbook()
    os2 = ow2.active
    os2.title = idt.NORMALIZED_SHEET
    os2.append(["Route"] + idt.SHARED_HEADER)
    os2.append(br2[0])
    ow2.save(old2)
    ow2.close()
    _a2, _bx2, sc32, note32 = eid.load_sides(a_cons2, old2)
    check("legacy library -> sidecar None + 'rebuild the TSN library' hint",
          sc32 is None and note32 and "rebuild the TSN library" in note32)
finally:
    import shutil
    shutil.rmtree(tmp2, ignore_errors=True)

# --------------------------------------------------------------------------- #
print("on-demand per-cell evidence (v0.23.0): the freshness gate")
import time                                                # noqa: E402
import matrix                                              # noqa: E402

tmp3 = Path(tempfile.mkdtemp())
try:
    store = tmp3 / "cell" / "intersection_detail_pdf"
    store.mkdir(parents=True)
    consolidated = matrix.consolidated_store_path(store, "intersection_detail_pdf")
    tsn = tmp3 / "tsn.xlsx"
    cmpwb = tmp3 / "cmp.xlsx"
    pdfdir = tmp3 / "pdfs"

    def _touch(p, when):
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"x")
        os.utime(p, (when, when))

    def _publish_comparison(p):
        import hashlib
        st = p.stat()
        digest = hashlib.sha256(p.read_bytes()).hexdigest()
        member = {
            "flavor": "values", "relative_path": p.name, "path": str(p),
            "canonical_path_at_write": str(p.resolve()),
            "commit_role": "canonical", "sha256": digest,
            "size": st.st_size, "mtime_ns": st.st_mtime_ns,
        }
        typed = ComparisonOutcome(
            status="ok", completion="complete", verdict="match",
            counts=ComparisonCounts(known=True, paired_rows=1),
            pairing_quality="exact")
        generation = ArtifactGeneration(
            generation_id="evidence-fixture", members=(member,),
            content_digests={"values": digest}, completion="complete",
            publication_state="committed", requested_mode="values")
        result = ConsolidateResult(
            status="ok", output_path=str(p), verdict="match",
            completion="complete", skipped_inputs=0, failed_inputs=0,
            comparison_outcome=typed, artifact_generation=generation)
        assert consolidation_meta.write_comparison_outcomes(result)

    def _publish_consolidation(p):
        assert consolidation_meta.write_outcome(
            p, ConsolidateResult(
                status="ok", output_path=str(p), completion="complete",
                skipped_inputs=0, failed_inputs=0))

    def _gate_error(expected_generation="evidence-fixture"):
        try:
            matrix.run_evidence_only("intersection_detail_pdf", store,
                                     "intersection_detail_pdf", tsn, cmpwb, pdfdir,
                                     events=None, examples=2,
                                     source_identity_check=lambda: True,
                                     expected_generation_id=expected_generation,
                                     source_workbook_identity=(
                                         tsn_library.normalized_workbook_identity(tsn)),
                                     live_tsn_path=tsn)
        except ValueError as e:
            return str(e)
        return None

    try:
        matrix.run_evidence_only("ramp_summary", store, "ramp_summary", tsn, cmpwb,
                                 pdfdir, events=None)
        _cap_err = None
    except ValueError as e:
        _cap_err = str(e)
    check("an evidence-incapable row is refused with the reason",
          _cap_err and "doesn't support evidence images" in _cap_err)

    _touch(tsn, time.time() - 200)
    err = _gate_error()
    check("missing comparison -> 'run the comparison first'",
          err and "run the comparison first" in err)

    now = time.time()
    _touch(cmpwb, now - 50)
    _publish_comparison(cmpwb)
    generation_err = _gate_error("wrong-generation")
    check("cache generation mismatch is refused before evidence rendering",
          generation_err and "cache generation do not match" in generation_err)
    err = _gate_error()
    check("missing consolidated -> 'run the comparison first'",
          err and "no consolidated" in err and "run the comparison first" in err)

    # a store file NEWER than the consolidated -> the store-changed refusal
    _touch(consolidated, now - 100)
    artifact_store.write_consolidated_fingerprint(consolidated, store)
    _publish_consolidation(consolidated)
    _touch(store / "intersection_detail_route_001.pdf", now - 20)
    err = _gate_error()
    check("store changed since the consolidation -> refuse with the refresh hint",
          err and "exports changed" in err and "refresh the comparison" in err)

    # consolidated fresh vs store, but NEWER than the comparison -> refuse.
    # (No fingerprint sidecar exists for this synthetic store; stub the staleness
    # probe so the mtime gates are what's under test.)
    _real_stale = matrix._consolidated_stale
    matrix._consolidated_stale = lambda *_a, **_k: False
    try:
        _touch(consolidated, now - 10)
        err = _gate_error()
        check("consolidated newer than the comparison -> refuse with the hint",
              err and "newer than" in err and "refresh the comparison" in err)

        _touch(consolidated, now - 100)
        _touch(tsn, now - 5)
        err = _gate_error()
        check("TSN workbook newer than the comparison -> refuse with the hint",
              err and "TSN workbook is newer" in err)

        # everything consistent -> the gate passes through to the generator; a
        # stubbed generate proves the call shape + the ok result + note.
        _touch(tsn, now - 200)
        import visual_evidence as _ve2
        _real_gen = _ve2.generate
        _ve2.generate = (lambda *_a, **_k:
                         {"note": "evidence: 2 example(s) across 1/1 …"})
        try:
            res = matrix.run_evidence_only(
                "intersection_detail_pdf", store, "intersection_detail_pdf", tsn, cmpwb,
                pdfdir, events=None, examples=2,
                source_identity_check=lambda: True,
                expected_generation_id="evidence-fixture",
                source_workbook_identity=(
                    tsn_library.normalized_workbook_identity(tsn)),
                live_tsn_path=tsn)
        finally:
            _ve2.generate = _real_gen
        check("fresh inputs -> ok result carrying the generator's note",
              res.status == "ok" and "example(s)" in (res.message or "")
              and res.summary_lines == [res.message])
    finally:
        matrix._consolidated_stale = _real_stale
finally:
    import shutil
    shutil.rmtree(tmp3, ignore_errors=True)

# --------------------------------------------------------------------------- #
print("Highway Log adapter (v0.24.0): fields, window map, routing, ditto discipline")
check("FIELDS = every Highway Log column except the Location key",
      ehl.FIELDS == [f for f in hlc.HEADER if f != hlc.HEADER[0]]
      and len(ehl.FIELDS) == 30)
check("field -> TSN window map is positional over ROW_KEYS and complete "
      "(Description alone has no window — its own follow-on lines)",
      ehl._TSN_WIN_KEY == dict(zip(hlc.HEADER, ctnl.ROW_KEYS))
      and all(f == "Description" or ehl._TSN_WIN_KEY[f] in ehl._TSN_WINDOWS
              for f in ehl.FIELDS))
check("verification projection == the comparator's load normalization + Excel TRIM "
      "(tab-padded values compare clean, numerics match the trim)",
      ehl.project("HG", "D\t\t") == "D"
      and ehl.project("Length (MI) [MI]", " 000.075 ") == "000.075")
check("canonical key: the comparator's roadbed_canonical_location (suffix "
      "authoritative; a dittoed LEFT block tags the row R)",
      ehl._canon(["012.887R"] + [None] * 30) == "012.887R"
      and ehl._canon(["012.887"] + [None] * 9 + ["+"] * 8 + [None] * 13) == "012.887R")
check("district_index is the sentinel single-folder entry (per-print routing)",
      ehl.district_index(Path("C:/anywhere")) == {"": Path("C:/anywhere")})
# Ditto discipline: a `+`-run cell on either side is NON-ASSERTING in the
# comparison, so enumerate_diffs must never sample it — while a genuine text
# diff in the same row still enumerates.
_hl_a = ["001"] + ["012.887"] + ["a"] * 30
_hl_b = ["001"] + ["012.887"] + ["a"] * 30
_hl_a[2], _hl_b[2] = "X", "+"                       # ditto side -> non-asserting
_hl_a[3], _hl_b[3] = "Y", "Z"                       # a real diff
_diffs = ehl.enumerate_diffs([_hl_a], [_hl_b], {"routing": "per-print"})
check("enumerate_diffs skips ditto cells but keeps real diffs (compared_cell semantics)",
      hlc.HEADER[1] not in _diffs and [e["key"] for e in _diffs[hlc.HEADER[2]]] == ["012.887"]
      and _diffs[hlc.HEADER[2]][0]["dist"] == "" and _diffs[hlc.HEADER[2]][0]["cnty"] == "")
check("enumerate_diffs judges through the LIVE schema (ditto_nonasserting set)",
      chl_cmp._SCHEMA.ditto_nonasserting is True)
# load_sides refuses per-route (route-less) workbooks: evidence groups by the
# leading Route column, which a per-route export doesn't carry.
_hl_tmp = Path(tempfile.mkdtemp(prefix="tsmis_ev_hl_"))
_wb = Workbook()
_ws = _wb.active
_ws.title = chl_cmp.SHEET_NAME
_ws.append(hlc.HEADER)                              # per-route: NO Route column
_ws.append(["012.887"] + ["a"] * 30)
_wb.save(_hl_tmp / "per_route.xlsx")
_r_t, _r_n, _sc, _note = ehl.load_sides(str(_hl_tmp / "per_route.xlsx"),
                                        str(_hl_tmp / "per_route.xlsx"))
check("load_sides refuses per-route (route-less) workbooks with a clear note",
      _sc is None and "Route column" in (_note or ""))
_wb2 = Workbook()
_ws2 = _wb2.active
_ws2.title = chl_cmp.SHEET_NAME
_ws2.append([hlc.ROUTE_COL] + hlc.HEADER)           # consolidated shape
_ws2.append(["001", "012.887"] + ["a"] * 30)
_wb2.save(_hl_tmp / "consolidated.xlsx")
_r_t, _r_n, _sc2, _note2 = ehl.load_sides(str(_hl_tmp / "consolidated.xlsx"),
                                          str(_hl_tmp / "consolidated.xlsx"))
check("load_sides accepts consolidated workbooks (truthy routing sidecar, no note)",
      _sc2 == {"routing": "per-print"} and _note2 is None
      and len(_r_t) == 1 and _r_t[0][0] == "001")
import shutil as _sh
_sh.rmtree(_hl_tmp, ignore_errors=True)

# --------------------------------------------------------------------------- #
print("Highway Sequence adapter (v0.25.0): fields, maps, routing, context discipline")
check("FIELDS = every shared column except the PM key",
      ehsl.FIELDS == [f for f in chsl_cmp.SHARED_HEADER if f != "PM"]
      and len(ehsl.FIELDS) == 6)
check("field -> TSMIS print column / TSN print window maps are complete",
      all(f in ehsl._TSMIS_COL for f in ehsl.FIELDS)
      and all(f == "Description" or f in ehsl._TSN_WIN for f in ehsl.FIELDS))
check("verification projection == the comparator's per-field normalization + TRIM "
      "(SIDE-AWARE Description per CMP-AUD-204: TSMIS strips its OWN-route "
      "label only, TSN is verbatim; county period)",
      ehsl.project("Description", "001/NB  OFF TO X ", route="001") == "NB OFF TO X"
      and ehsl.project("Description", "1/103 SEP", route="680") == "1/103 SEP"
      and ehsl.project("Description", "1/103 SEP", side="tsn") == "1/103 SEP"
      and ehsl.project("County", "LA.") == "LA"
      and ehsl.project("FT", " H\t") == "H")
check("canonical key: 'COUNTY GLUED-POSTMILE', county normalized",
      ehsl._canon("LA.", "R000.129") == "LA R000.129"
      and ehsl._canon("ORA", "018.530E") == "ORA 018.530E")
check("district_index is the sentinel single-folder entry (per-print routing)",
      ehsl.district_index(Path("C:/anywhere")) == {"": Path("C:/anywhere")})
# Context discipline: HSL has NO context fields since 2026-08-10 (owner: compare
# HG / City / Distance too), so enumerate_diffs must now sample all five
# differing columns. That is also the proof it reads the LIVE schema rather than
# a hardcoded skip list — a stale list would drop exactly those three.
_hs_a = ["001", "ORA", "R000.129", "LGNB", "D", "H", "000.100", "JCT 5"]
_hs_b = ["001", "ORA", "R000.129", "",     "U", "I", "000.900", "JCT 5 UC"]
_hs_diffs = ehsl.enumerate_diffs([_hs_a], [_hs_b], {"routing": "per-print"})
check("enumerate_diffs samples every differing column, context set now empty",
      set(_hs_diffs) == {"City", "HG", "FT", "Distance To Next Point",
                         "Description"}
      and [e["key"] for e in _hs_diffs["FT"]] == ["ORA R000.129"]
      and _hs_diffs["FT"][0]["dist"] == "" and _hs_diffs["FT"][0]["cnty"] == "")
check("enumerate_diffs judges through the LIVE schema, not a hardcoded list",
      tuple(chsl_cmp._SCHEMA.context_fields) == ()
      and {"HG", "City", "Distance To Next Point"} <= set(_hs_diffs))
# LOCKSTEP pins vs the PDF consolidator: the wrap join, the PM-less data test,
# the trailer heading, and the evidence twin of the word classifier.
check("join_desc_parts: bare after a hyphen, one space otherwise, empties skipped",
      chslp.join_desc_parts(["UC 55-", "1107"]) == "UC 55-1107"
      and chslp.join_desc_parts(["A", "", "B"]) == "A B"
      and chslp.join_desc_parts(["", "X"]) == "X")
check("PM-less data rows accepted (END OF ROUTE / CITY END), furniture rejected",
      chslp._is_pmless_data({"pm": "", "prefix": "", "suffix": "",
                             "desc": "END OF ROUTE 043", "hg": "D", "ft": "H",
                             "county": "", "city": "", "dist": "000.000"})
      and not chslp._is_pmless_data({"pm": "", "prefix": "", "suffix": "Direction:",
                                     "desc": "S - N", "hg": "", "ft": "",
                                     "county": "", "city": "", "dist": ""})
      and not chslp._is_pmless_data({"pm": "", "prefix": "", "suffix": "",
                                     "desc": "", "hg": "D", "ft": "H",
                                     "county": "", "city": "", "dist": ""}))
check("the trailer heading pin (parsing hard-stops there)",
      chslp.TRAILER_HEADING == "Unresolved Intersections")
# The evidence classifier is the word-object-keeping TWIN of the consolidator's:
# the same synthetic line must classify identically through both.
_hd7 = {"COUNTY": {"x0": 31, "x1": 68}, "CITY": {"x0": 84, "x1": 103},
        "PM": {"x0": 149, "x1": 163}, "HG": {"x0": 201, "x1": 214},
        "FT": {"x0": 225, "x1": 235}, "NEXT": {"x0": 251, "x1": 274},
        "DESCRIPTION": {"x0": 317, "x1": 376}}
_bounds = chslp._boundaries(_hd7)
_line = [{"text": "ORA", "x0": 40, "x1": 59}, {"text": "LGNB", "x0": 82, "x1": 105},
         {"text": "R", "x0": 127, "x1": 133}, {"text": "000.129", "x0": 140, "x1": 173},
         {"text": "E", "x0": 184, "x1": 189}, {"text": "D", "x0": 204, "x1": 211},
         {"text": "H", "x0": 233, "x1": 239}, {"text": "000.124", "x0": 262, "x1": 294},
         {"text": "COUNTY", "x0": 317, "x1": 353}, {"text": "BEGIN:", "x0": 356, "x1": 384}]
_vals = chslp._classify_words(_line, _bounds)
_cols = ehsl._classify_line_words(_line, _bounds)
check("consolidator + evidence classify one line identically (LOCKSTEP twin)",
      _vals == {k: " ".join(w["text"] for w in ws) for k, ws in _cols.items()}
      and _vals["county"] == "ORA" and _vals["prefix"] == "R"
      and _vals["pm"] == "000.129" and _vals["suffix"] == "E"
      and _vals["hg"] == "D" and _vals["ft"] == "H"
      and _vals["dist"] == "000.124" and _vals["desc"] == "COUNTY BEGIN:")
# load_sides refuses a NON-consolidated TSMIS workbook (no Route column) with the
# comparator's own hint, and accepts the consolidated + normalized-TSN pair.
_hs_tmp = Path(tempfile.mkdtemp(prefix="tsmis_ev_hsl_"))
_wbp = Workbook()
_wsp = _wbp.active
_wsp.title = chsl_cmp.TSMIS_SHEET
_wsp.append(["County", "City", None, "PM", None, "HG", "FT",
             "Distance To Next Point", "Description"])   # per-route: NO Route col
_wsp.append(["ORA", None, "R", "000.129", None, "D", "H", "000.124", "X"])
_wbp.save(_hs_tmp / "per_route.xlsx")
_wbn = Workbook()
_wsn = _wbn.active
_wsn.title = ctnsl.NORMALIZED_SHEET
_wsn.append(ctnsl.NORMALIZED_HEADER)
_wsn.append(["001", "ORA", "R000.129", None, "D", "H", "000.102", "X"])
_mkn = _wbn.create_sheet(ctnsl.MARKER_SHEET)          # the v4 shape marker
_mkn.append(["Report", ctnsl.REPORT_NAME])
_mkn.append(["Normalization version", ctnsl.NORMALIZATION_VERSION])
_wbn.save(_hs_tmp / "tsn.xlsx")
_r_t, _r_n, _sc3, _note3 = ehsl.load_sides(str(_hs_tmp / "per_route.xlsx"),
                                           str(_hs_tmp / "tsn.xlsx"))
check("load_sides refuses per-route (route-less) workbooks with a clear note",
      _sc3 is None and "consolidate a fresh export" in (_note3 or ""))
_wbc = Workbook()
_wsc = _wbc.active
_wsc.title = chsl_cmp.TSMIS_SHEET
_wsc.append(["Route", "County", "City", None, "PM", None, "HG", "FT",
             "Distance To Next Point", "Description"])    # consolidated shape
_wsc.append(["001", "ORA", None, "R", "000.129", None, "D", "H", "000.124", "X"])
_wbc.save(_hs_tmp / "consolidated.xlsx")
_r_t, _r_n, _sc4, _note4 = ehsl.load_sides(str(_hs_tmp / "consolidated.xlsx"),
                                           str(_hs_tmp / "tsn.xlsx"))
check("load_sides accepts the consolidated + normalized-TSN pair (typed glued "
      "PhysicalKey both sides, CMP-AUD-045)",
      _sc4 == {"routing": "per-print"} and _note4 is None
      and len(_r_t) == 1 and _r_t[0][0] == "001" and str(_r_t[0][2]) == "R000.129"
      and dict(_r_t[0][2].physical_identity.canonical_components)["postmile"]
      == "R000.129"
      and len(_r_n) == 1 and str(_r_n[0][2]) == "R000.129"
      and _r_t[0][2] == _r_n[0][2])
# a pre-v4 normalized workbook (no marker sheet) is refused with the rebuild note
_wbo = Workbook()
_wso = _wbo.active
_wso.title = ctnsl.NORMALIZED_SHEET
_wso.append(ctnsl.NORMALIZED_HEADER)
_wso.append(["001", "ORA", "R000.129", None, "D", "H", "000.102", "X"])
_wbo.save(_hs_tmp / "tsn_old.xlsx")
_r_t, _r_n, _sc5, _note5 = ehsl.load_sides(str(_hs_tmp / "consolidated.xlsx"),
                                           str(_hs_tmp / "tsn_old.xlsx"))
check("load_sides refuses a pre-v4 normalized TSN workbook with the rebuild note",
      _sc5 is None and "older TSN Highway Sequence converter" in (_note5 or ""))
_sh.rmtree(_hs_tmp, ignore_errors=True)

# --------------------------------------------------------------------------- #
print("Ramp Detail adapter (v0.26.0): fields, maps, projections, dual-row discipline")
check("FIELDS = the union of both flavors' compared columns (PM key + the two "
      "always-context TSN columns excluded; District joined in CMP-AUD-185)",
      erd.FIELDS == ["PR", "District", "Date of Record", "HG", "Area 4",
                     "City Code", "R/U", "Description", "On/Off", "Ramp Type"])
check("field -> TSMIS print column / TSN print window maps are complete",
      all(f in erd._TSMIS_COL for f in erd.FIELDS)
      and all(f in erd.TSN_CELL for f in erd.FIELDS)
      and set(erd._TSMIS_COL.values()) <= set(crdpdf._COL_ORDER) | {"loc"}
      and all(n in {w[0] for w in erd._L_WIN} for n in erd.TSN_CELL.values()))
check("TSN windows are x-ordered and non-overlapping (the fixed template)",
      all(erd._L_WIN[i][2] <= erd._L_WIN[i + 1][1]
          for i in range(len(erd._L_WIN) - 1)))
check("verification projection == the PDF flavor's per-field normalization + TRIM "
      "(route-prefix strip + collapse + the null-render message on Description; "
      "the '-' null marks; the print's N -> TSN's O)",
      erd.project("Description", "001/NB  OFF TO X ") == "NB OFF TO X"
      and erd.project("Description", "NO RAMP LINEAR EVENT") == ""
      and erd.project("Area 4", "-") == ""
      and erd.project("On/Off", "-") == ""
      and erd.project("On/Off", "N") == "O"
      and erd.project("On/Off", "F") == "F"
      and erd.project("Date of Record", "02/25/1976") == "1976-02-25"
      and erd.project("Ramp Type", " D ") == "D")
# Dual-row discipline: the Excel row's comparison keeps On/Off + Ramp Type as
# context (never enumerated); the PDF row's comparison COMPARES them. Ramp Name
# and ADT never enumerate on either row.
# The PM key cells carry the D4 PhysicalKey (CMP-AUD-045); District is a
# compared column, equal here so it never enumerates.
_rd_k = crd_cmp._physical_pm_key("001", "ORA", "000.606",
                                 (("route", "001"),), "fixture")
_rd_a = ["001", "R", _rd_k, "12", "1976-02-25", "D", "Y", "DAPT", "U",
         "NB OFF X", "", "O", "D", ""]
_rd_b = ["001", "M", crd_cmp._physical_pm_key("001", "ORA", "000.606",
                                              (("route", "001"),), "fixture"),
         "12", "1976-02-25", "L", "Y", "DAPT", "U",
         "NB OFF Y", "RAMP NM", "F", "L", "070"]
_dc = {("001", _rd_k): [("12", "ORA")]}
_dx = erd.enumerate_diffs([_rd_a], [_rd_b], {"dc": _dc, "pdf": False})
check("Excel-row enumerate_diffs skips the print-only + TSN-only columns",
      set(_dx) == {"PR", "HG", "Description"}
      and _dx["PR"][0]["dist"] == "12" and _dx["PR"][0]["cnty"] == "ORA")
_dp = erd.enumerate_diffs([_rd_a], [_rd_b], {"dc": _dc, "pdf": True})
check("PDF-row enumerate_diffs ALSO samples On/Off + Ramp Type (compared there)",
      set(_dp) == {"PR", "HG", "Description", "On/Off", "Ramp Type"})
check("enumerate_diffs judges through the LIVE schemas (context sets)",
      set(crd_cmp._SCHEMA.context_fields)
      == {"Ramp Name", "On/Off", "Ramp Type", "ADT"}
      and set(crdp.TSMIS_PDF_VS_TSN._schema.context_fields)
      == {"Ramp Name", "ADT"})
check("_schema_for binds each flavor to the live comparator context (CMP-AUD-107)",
      erd._schema_for(False) is crd_cmp._SCHEMA
      and set(erd._schema_for(True).context_fields)
      == set(crdp.TSMIS_PDF_VS_TSN._schema.context_fields))
# LOCKSTEP pins vs the PDF consolidator: the wrap join, the PM test, the prefix
# legend, and the null-render tokens the projections must keep matching.
check("join_desc_parts: bare after a hyphen, one space otherwise, empties skipped",
      crdpdf.join_desc_parts(["UC 55-", "1107"]) == "UC 55-1107"
      and crdpdf.join_desc_parts(["A", "", "B"]) == "A B")
check("the PM row test + prefix legend pins",
      crdpdf.PM_RE.fullmatch("000.606") and not crdpdf.PM_RE.fullmatch("0.606")
      and crdpdf.PREFIX_SET == frozenset("CDGHLMNRST"))
check("the null-render token pins (the comparison flavors project these)",
      crdp._NULL_DESC == "NO RAMP LINEAR EVENT" and crdp._NULL_MARK == "-")
# The consolidated-workbook source detector: the PDF-consolidated carries the
# print-only "On/Off" header; the Excel-consolidated doesn't.
_rd_tmp = Path(tempfile.mkdtemp(prefix="tsmis_ev_rd_"))
_wbe = Workbook()
_wse = _wbe.active
_wse.title = crd_cmp.TSMIS_SHEET
_wse.append(["Route", "Location", None, "PM", "Date of Record", None, "HG",
             "Area 4", None, "City Code", "R/U", "Description"])
_wse.append(["001", "12-ORA-001", "R", "000.606", "02/25/1976", None, "D", "Y",
             "DAPT", "U", "001/NB OFF X", None])
_wbe.save(_rd_tmp / "excel_cons.xlsx")
_wbP = Workbook()
_wsP = _wbP.active
_wsP.title = crd_cmp.TSMIS_SHEET
_wsP.append(["Route"] + list(crdpdf.HEADER))
_wsP.append(["001", "12-ORA-001", "R", "000.606", "02/25/1976", None, "D", "Y",
             "DAPT", "U", "001/NB OFF X", None, "N", "D"])
_wbP.save(_rd_tmp / "pdf_cons.xlsx")
check("_is_pdf_consolidated tells the two consolidated shapes apart",
      erd._is_pdf_consolidated(str(_rd_tmp / "pdf_cons.xlsx"))
      and not erd._is_pdf_consolidated(str(_rd_tmp / "excel_cons.xlsx")))
# load_sides on a sidecar-less normalized TSN library returns the rebuild note
# instead of guessing (the pre-v3 library case).
_wbn = Workbook()
_wsn = _wbn.active
_wsn.title = crd_cmp.NORMALIZED_SHEET
_wsn.append(["Route"] + [h for h in crd_cmp.SHARED_HEADER if h != "District"]
            + ["TSN District", "TSN County"])           # v3: pre-county identity
_wsn.append(["001", "R", "000.606", "1976-02-25", "D", "Y", "DAPT", "U",
             "NB OFF X", "", "O", "D", "", "12", "ORA"])
_wbn.save(_rd_tmp / "tsn_v2.xlsx")
_t_r, _t_n, _meta, _note = erd.load_sides(str(_rd_tmp / "pdf_cons.xlsx"),
                                          str(_rd_tmp / "tsn_v2.xlsx"))
check("load_sides refuses a sidecar-less (pre-v3) TSN library with the rebuild note",
      _meta is None and "rebuild the TSN library" in (_note or ""))
_sh.rmtree(_rd_tmp, ignore_errors=True)

# --------------------------------------------------------------------------- #
print("engine misc")
check("reason summarizer dedupes and caps",
      ve._summarize_reasons(["a", "a", "b", "c", "d"]) == "a; b; c"
      and ve._summarize_reasons([]) == "no candidates")
check("evidence never keys off visible text (regex sanity: safe filename)",
      re.sub(r"[^A-Za-z0-9]+", "_", "Med V/WDA").strip("_") == "Med_V_WDA")

# ----------------------------------------------------------------------------- #
print("the pdf/ drop folder exists for the user (v0.21.1 — the update-day gap)")
import paths                                              # noqa: E402
import report_catalog                                     # noqa: E402
_pdf_drop_reports = set(ve.TSN_PDF_REPORT.values()) - ve._TSN_PDFS_IN_RAW
check("every pdf/-drop TSN source is catalog-flagged evidence_pdfs (the "
      "catalog also keeps highway_detail's flagged drop folder — the reserved "
      "fifth family, pre-release, not in the evidence registry)",
      {report_catalog.TSN[[e.subdir for e in report_catalog.TSN].index(r)].evidence_pdfs
       for r in _pdf_drop_reports} == {True}
      and {e.subdir for e in report_catalog.TSN if e.evidence_pdfs}
      == _pdf_drop_reports | {"highway_detail"})
check("every raw-sourced evidence report is a district_pdfs TSN library (its "
      "prints ARE the raw inputs, so no pdf/ drop folder is flagged)",
      all(report_catalog.TSN[[e.subdir for e in report_catalog.TSN].index(r)].raw_kind
          == "district_pdfs"
          and not report_catalog.TSN[[e.subdir for e in report_catalog.TSN]
                                     .index(r)].evidence_pdfs
          for r in ve._TSN_PDFS_IN_RAW)
      and ve._TSN_PDFS_IN_RAW <= set(ve.TSN_PDF_REPORT.values()))
_tmp = Path(tempfile.mkdtemp())
_old_root = paths.TSN_LIBRARY_ROOT
try:
    paths.TSN_LIBRARY_ROOT = _tmp / "tsn_library"
    root = tsn_library.ensure_layout()
    pdf = root / "highway_detail" / "pdf"
    check("ensure_layout creates highway_detail/pdf/ + drops the hint",
          pdf.is_dir() and any(pdf.glob("_PUT TSN DISTRICT PDFS HERE.txt")))
    check("…and the pdf/ path == the library's pdf_dir (the engine no longer "
          "maps highway_detail — the reserved pre-release family)",
          pdf == tsn_library.pdf_dir("highway_detail"))
    ipdf = root / "intersection_detail" / "pdf"
    check("ensure_layout creates intersection_detail/pdf/ + its hint (v0.22.0)",
          ipdf.is_dir() and any(ipdf.glob("_PUT TSN DISTRICT PDFS HERE.txt")))
    check("…and it too == the engine's tsn_pdf_dir",
          ipdf == ve.tsn_pdf_dir("intersection_detail_pdf")
          == tsn_library.pdf_dir("intersection_detail"))
    rpdf = root / "ramp_detail" / "pdf"
    check("ensure_layout creates ramp_detail/pdf/ + its hint (v0.26.0)",
          rpdf.is_dir() and any(rpdf.glob("_PUT TSN DISTRICT PDFS HERE.txt")))
    check("…and it too == the engine's tsn_pdf_dir",
          rpdf == ve.tsn_pdf_dir("ramp_detail_pdf")
          == tsn_library.pdf_dir("ramp_detail"))
    readme = root / tsn_library._README_NAME
    check("the root README documents BOTH pdf/ folders",
          readme.is_file()
          and "highway_detail/pdf/" in readme.read_text(encoding="utf-8")
          and "intersection_detail/pdf/" in readme.read_text(encoding="utf-8"))
    # an OUTDATED readme (an updated install) refreshes on the next launch
    readme.write_text("old text from a previous version\n", encoding="utf-8")
    tsn_library.ensure_layout()
    check("a stale README from an older install is refreshed",
          "highway_detail/pdf/" in readme.read_text(encoding="utf-8"))
finally:
    paths.TSN_LIBRARY_ROOT = _old_root
    import shutil as _sh
    _sh.rmtree(_tmp, ignore_errors=True)

# --------------------------------------------------------------------------- #
# HF-05 (PCOA-FINAL-005): blank-side target geometry never guesses. A field
# with no cell rectangle inside the record's own printed lines REFUSES —
# below-the-record and fixed-zone fallbacks boxed the NEXT record / the final
# 'O' of 'EQUATES TO' in the audited sets.
print("HF-05: blank-side targets refuse instead of guessing")
import evidence_highway_log as _ehl5
import evidence_highway_sequence as _ehs5

_hl_rec = {"approx": False, "page": 3, "top": 100.0, "bottom": 110.0,
           "src": "x.pdf",
           "chars": [{"x0": 60.0, "x1": 70.0, "top": 100.0, "bottom": 110.0,
                      "text": "R"}],
           "windows": [(0, 100)] * 30, "desc": [],
           "row": ["R000.100"] + [""] * 30}
check("HL TSMIS blank Description -> no geometry (was: a box BELOW the record)",
      _ehl5.tsmis_box(_hl_rec, "Description") is None)
# A TRAILING blank column on a short row: the column's window begins PAST the
# line's last glyph, so there is no cell rectangle inside the record's own
# printed extent. The old fallback drew a degenerate 10-point box out in the
# inter-column whitespace beyond the record (the RB4-A1 native-scale
# inspection measured one 8 px wide, sitting outside the record outline).
_hl_trail = dict(_hl_rec, windows=[(0, 100)] * 29 + [(400.0, 500.0)])
check("HL TSMIS trailing blank column -> no geometry, never a box past the "
      "record's own printed extent",
      _ehl5.tsmis_box(_hl_trail, hlc.HEADER[-1]) is None)
# A window the record's ink only just REACHES is the same defect wearing a
# smaller size: refusing only an EMPTY overlap left a record whose last glyph
# poked a point or two in (or whose cosmetic padding did) still drawing a
# sliver in the gap after the PREVIOUS column — two RB4-A1 inspectors measured
# one at 12 px, sitting past the record's right edge and well left of where
# the column actually prints. A blank cell is only boxable when the record's
# own ink BRACKETS the window on both sides.
_hl_graze = dict(_hl_rec,
                 chars=[{"x0": 60.0, "x1": 405.0, "top": 100.0,
                         "bottom": 110.0, "text": "R"}],
                 windows=[(0, 100)] * 29 + [(400.0, 500.0)])
check("HL TSMIS blank column the record's ink merely GRAZES -> no geometry "
      "(the sliver case, not just the empty-overlap case)",
      _ehl5.tsmis_box(_hl_graze, hlc.HEADER[-1]) is None)
# ...while a blank column the record's ink BRACKETS still boxes, so the
# refusal is targeted rather than a blanket ban on blank cells.
_hl_mid = dict(_hl_rec,
               chars=[{"x0": 10.0, "x1": 600.0, "top": 100.0, "bottom": 110.0,
                       "text": "R"}],
               windows=[(0, 100)] * 29 + [(400.0, 500.0)])
_hl_mid_box = _ehl5.tsmis_box(_hl_mid, hlc.HEADER[-1])
check("...but a blank column the record BRACKETS still gets its box, spanning "
      "the whole column window",
      _hl_mid_box is not None
      and _hl_mid_box[1][0] <= 400.0 and _hl_mid_box[1][2] >= 500.0)
# The env lane rides the same geometry (env_box delegates to tsmis_box), and
# the third measured sliver was found THERE — on a blank Date of Rec, not on
# the vs-TSN side. Pin the delegation so env can never grow its own rule.
check("HL env blank geometry IS the TSMIS rule (env_box delegates), so the "
      "graze case refuses on the env lane too",
      _ehl5.env_box(_hl_graze, hlc.HEADER[-1]) is None
      and _ehl5.env_box(_hl_mid, hlc.HEADER[-1]) is not None)
# RD District is DERIVED from the composite Location cell, and `cols["loc"]`
# is the bucketing loop's CATCH-ALL for every unwindowed word left of the
# Description — the postmile has no window of its own and lands there too.
# Boxing the bucket swallowed the P/R/E and PM columns' values; two
# independent RB4-A1 inspectors measured it against the print's own ruled
# columns. The box must be the LOCATION cell alone.
def _rd_word(text, x0, x1):
    return {"text": text, "x0": x0, "x1": x1, "top": 10.0, "bottom": 18.0}


_rd_loc = _rd_word("12-SD-005", 30.0, 100.0)
_rd_pm = _rd_word("072.366", 140.0, 185.0)      # unwindowed -> the loc bucket
_rd_geo = {"page": 2, "top": 10.0, "bottom": 18.0,
           "cols": {"loc": [_rd_loc, _rd_pm]},
           "b": {"loc_pr": 111.0, "type_desc": 500.0},
           "words": [_rd_loc, _rd_pm]}
_rd_box = erd.tsmis_box(_rd_geo, "District")
check("RD District boxes the LOCATION cell alone, never the catch-all bucket "
      "(which carries the postmile too)",
      _rd_box is not None and _rd_box[1][0] <= _rd_loc["x0"]
      and _rd_box[1][2] < _rd_pm["x0"])
check("...and a record with nothing in the Location cell refuses rather than "
      "boxing whatever else landed in the bucket",
      erd.tsmis_box(dict(_rd_geo, cols={"loc": [_rd_pm]}), "District") is None)
_hl_tsn_rec = {"page": 2, "top": 50.0, "bottom": 58.0, "src": "d.pdf",
               "chars": [{"x0": 10.0, "x1": 30.0, "top": 50.0, "bottom": 58.0,
                          "text": "T121.831"}],
               "desc": [], "rowd": {"description": None}}
check("HL TSN blank Description -> no geometry (was: the NEXT record's box)",
      _ehl5.tsn_box(_hl_tsn_rec, "Description") is None)
# The TSN side's own trailing-blank column. `_line_cell_box` learned to refuse
# this on the TSMIS side; the TSN side kept a `lo, lo + 10` fallback, so three
# RB4-A1 inspectors independently measured a ~7 px sliver beyond the record's
# printing on Sig Chg. Date. Both sides must refuse, and both must still box a
# blank column that DOES overlap the record.
import consolidate_tsn_highway_log as _ctnl5

_sig_lo = _ehl5._TSN_WINDOWS[_ehl5._TSN_WIN_KEY["Sig Chg. Date"]][0]


def _hl_tsn_chars(x1_end):
    return [{"x0": x1_end - 20.0, "x1": x1_end, "top": 50.0, "bottom": 58.0,
             "text": "6"}]


_hl_tsn_trail = dict(_hl_tsn_rec, chars=_hl_tsn_chars(_sig_lo - 40.0))
check("HL TSN trailing blank column -> no geometry, never a degenerate box "
      "past the record's own printed extent",
      _ehl5.tsn_box(_hl_tsn_trail, "Sig Chg. Date") is None)
# The measured field defect: the record's last token ends a couple of points
# INSIDE the window, so the overlap was non-empty and a sliver was drawn.
_hl_tsn_graze = dict(_hl_tsn_rec, chars=_hl_tsn_chars(_sig_lo + 2.0))
check("HL TSN blank column the record's ink merely GRAZES -> no geometry "
      "(the 12 px Sig Chg. Date sliver two inspectors measured)",
      _ehl5.tsn_box(_hl_tsn_graze, "Sig Chg. Date") is None)
# Sig Chg. Date is the LAST column, so nothing can ever print to its right —
# a blank there is never locatable from the record's own line. A bracketed
# interior column still boxes.
_mi_lo, _mi_hi = _ehl5._TSN_WINDOWS[_ehl5._TSN_WIN_KEY["Length (MI) [MI]"]]
_hl_tsn_mid = dict(_hl_tsn_rec,
                   chars=[{"x0": _mi_lo - 30.0, "x1": _mi_hi + 200.0,
                           "top": 50.0, "bottom": 58.0, "text": "6"}])
_hl_tsn_mid_box = _ehl5.tsn_box(_hl_tsn_mid, "Length (MI) [MI]")
check("...but a TSN blank column the record BRACKETS still gets its box, "
      "spanning the whole column window",
      _hl_tsn_mid_box is not None
      and _hl_tsn_mid_box[1][0] <= _mi_lo and _hl_tsn_mid_box[1][2] >= _mi_hi)
# ...and the raw-token hook must read the PRINT's token, not the value
# `_normalize_row` rewrote in place. Driven through the SHIPPED scan (a
# monkeypatched pdfplumber, like check_intersection_detail_pdf) so moving the
# snapshot back below the normalizer fails here — the defect was at the CALL
# SITE, not in the hook.
_MI_RAW, _TW_RAW = "0.071", "024"


class _FakeTsnPdf:
    def __init__(self, pages):
        self.pages = pages

    def __enter__(self):
        return self

    def __exit__(self, *a):
        return False


def _tsn_line(top, items, cw=2.4):
    out = []
    for center, text in items:
        x = center - (len(text) * cw) / 2
        for ch in text:
            if ch.strip():
                out.append({"text": ch, "x0": x, "x1": x + cw, "top": top,
                            "bottom": top + 7.0})
            x += cw
    return out


def _win_center(key):
    lo, hi = next((l, h) for k, l, h in _ctnl5.COLUMN_WINDOWS if k == key)
    return (lo + hi) / 2


def _scan_one_tsn_row():
    """Run the REAL _scan_tsn_print over one synthetic district print."""
    from types import SimpleNamespace
    chars = (_tsn_line(70.0, [(_win_center("location"), "012.540")])
             + _tsn_line(70.0, [(_win_center("mi"), _MI_RAW)])
             + _tsn_line(70.0, [(_win_center("rb_tw"), _TW_RAW)]))
    group = _tsn_line(60.0, [(260.0, "04"), (275.0, "ALA"), (292.0, "001")])
    page = SimpleNamespace(width=612.0, chars=group + chars, rects=[])
    found = __import__("collections").defaultdict(list)
    saved = _ehl5.pdfplumber.open

    # The scan keeps only rows whose canonical key is WANTED; this stands in for
    # the real needed-keys set so the synthetic row survives to the record.
    class _All(dict):
        def __contains__(self, _k):
            return True

    try:
        _ehl5.pdfplumber.open = lambda p: _FakeTsnPdf([page])
        _ehl5._scan_tsn_print(Path("D04 fake.pdf"), {"001"}, _All(), found)
    finally:
        _ehl5.pdfplumber.open = saved
    recs = [r for v in found.values() for r in v]
    return recs[0] if recs else None


_tsn_rec = _scan_one_tsn_row()
check("the shipped TSN scan parsed the synthetic row", _tsn_rec is not None)
if _tsn_rec is not None:
    check("HL TSN raw token is the PRINT's MI, not the zero-padded compared "
          "value (_normalize_row rewrites rowd IN PLACE)",
          _ehl5.tsn_raw(_tsn_rec, "Length (MI) [MI]") == _MI_RAW
          and _tsn_rec["rowd"]["mi"] != _MI_RAW)
    check("HL TSN raw token is the PRINT's traveled-way width, not the "
          "leading-zero-stripped compared value",
          _ehl5.tsn_raw(_tsn_rec, "RB T-W Wid [RB T-W]") == _TW_RAW
          and _tsn_rec["rowd"]["rb_tw"] != _TW_RAW)
    check("...so the engine DOES disclose those two forms on the image",
          ve._normalization_note(
              (None, _ehl5.tsn_raw(_tsn_rec, "RB T-W Wid [RB T-W]")),
              (None, _ehl5.tsn_value(_tsn_rec, "RB T-W Wid [RB T-W]")),
              ("TSMIS", "TSN")).startswith("TSN's boxed cell holds '024'"))
_hsl_eq = {"rowd": {"county": "ALA", "pm": "006.798", "city": None, "hg": None,
                    "ft": None, "dist": None, "description": "EQUATES TO"},
           "src": "d.pdf", "dist": "04", "cnty": "ALA", "page": 5,
           "equate": True, "top": 200.0, "bottom": 208.0,
           "words": [{"x0": 98.0, "x1": 130.0, "top": 200.0, "bottom": 208.0,
                      "text": "006.798"},
                     {"x0": 170.0, "x1": 205.0, "top": 200.0, "bottom": 208.0,
                      "text": "EQUATES"},
                     {"x0": 207.0, "x1": 220.0, "top": 200.0, "bottom": 208.0,
                      "text": "TO"}],
           "desc": []}
check("HSL equate line: FT/HG/City/Distance refuse (was: the final 'O' of "
      "'EQUATES TO')",
      _ehs5.tsn_box(_hsl_eq, "FT") is None
      and _ehs5.tsn_box(_hsl_eq, "HG") is None
      and _ehs5.tsn_box(_hsl_eq, "Distance To Next Point") is None)
check("HSL equate line: the synthesized Description has no printed segs -> "
      "no geometry",
      _ehs5.tsn_box(_hsl_eq, "Description") is None)
check("HSL equate line: County and PM (really printed) still box",
      _ehs5.tsn_box(_hsl_eq, "PM") is not None)

# The ENGINE backstop: whatever an adapter returns, an env target outside the
# captioned record's own printed lines is refused before rendering.
print("HF-10: the engine's row-rectangle backstop on env renders")


class _BadBoxAdapter:
    @staticmethod
    def env_value(_rec, _field):
        return "V"

    @staticmethod
    def env_box(_rec, _field):
        return (1, (10, 120, 30, 130), (100, 110), (0, 200))  # below the record


_bad_loc = ({"001": {"K": [{"src": "missing.pdf"}]}},
            {"001": {"K": [{"src": "missing.pdf"}]}})
_got, _reason = ve._env_example_sides(
    _BadBoxAdapter, {"route": "001", "key": "K", "display": "V ≠ W"},
    "F", _bad_loc, ("A", "B"), {})
check("an env target outside the record's own lines is refused by the ENGINE",
      _got is None and "outside the record's own printed lines" in _reason)


class _WideBoxAdapter:
    """Inside the record's LINES, but past its printed width — the axis the
    backstop used to ignore, which is exactly how a mis-anchored blank-cell
    window reached a render."""

    @staticmethod
    def env_value(_rec, _field):
        return "V"

    @staticmethod
    def env_box(_rec, _field):
        return (1, (180, 100, 260, 110), (100, 110), (0, 200))


_got, _reason = ve._env_example_sides(
    _WideBoxAdapter, {"route": "001", "key": "K", "display": "V ≠ W"},
    "F", _bad_loc, ("A", "B"), {})
check("an env target outside the record's printed WIDTH is refused too",
      _got is None and "outside the record's own printed width" in _reason)


class _EdgePadBoxAdapter:
    """A blank-cell window legitimately padded a few points past the record's
    glyph extent must still render — the width backstop must not be so tight
    that it refuses the honest case."""

    @staticmethod
    def env_value(_rec, _field):
        return "V"

    @staticmethod
    def env_box(_rec, _field):
        return (1, (-4, 100, 204, 110), (100, 110), (0, 200))


_got, _reason = ve._env_example_sides(
    _EdgePadBoxAdapter, {"route": "001", "key": "K", "display": "V ≠ W"},
    "F", _bad_loc, ("A", "B"), {})
check("...but a few points of legitimate edge padding is not refused for "
      "width (it fails later, on the unreadable print)",
      _got is None and "printed width" not in _reason)

# --------------------------------------------------------------------------- #
# PCOA-FINAL-005's other half, and the one the engine backstop CANNOT catch: a
# blank cell's rectangle may sit well inside the record's own lines and width
# and still enclose the NEIGHBOURING column's glyphs. The audit measured the
# previous fixed `boundary + offset … +30pt` guesses doing exactly that on
# thousands of Highway Sequence rows — a blank '(col C)' boxing the PM value, a
# blank '(col E)' boxing the HG letter.
# Resolving WHICH print a route's evidence comes from. The adapter's own name
# only exists under a run folder; an Export Everything store tags each name
# with its environment, so the engine falls back to the end-anchored
# `…_route_<token>.<ext>` contract. That fallback must never GUESS.
print("HF-10: the per-route print is resolved exactly or not at all")
_pr_dir = Path(tempfile.mkdtemp(prefix="check_ev_route_"))


class _PrintAdapter:
    @staticmethod
    def tsmis_pdf_path(pdf_dir, route):
        return Path(pdf_dir) / f"fixt_route_{route}.pdf"


def _resolve(route):
    return ve.find_route_print(_pr_dir, _PrintAdapter, route)


try:
    check("no candidate at all resolves to nothing", _resolve("001") is None)
    (_pr_dir / "ssor-prod fixt_route_001.pdf").write_bytes(b"%PDF-1.4\n")
    check("a store-tagged name resolves by the end-anchored route contract",
          _resolve("001") == _pr_dir / "ssor-prod fixt_route_001.pdf")
    (_pr_dir / "fixt_route_002.pdf").write_bytes(b"%PDF-1.4\n")
    check("another route's print is not borrowed", _resolve("003") is None)
    check("the exact adapter name wins when it exists",
          _resolve("002") == _pr_dir / "fixt_route_002.pdf")
    (_pr_dir / "ars-prod fixt_route_001.pdf").write_bytes(b"%PDF-1.4\n")
    check("TWO names honour the contract -> refuse, never pick one "
          "(captioning a route from the wrong environment's print)",
          _resolve("001") is None)
    (_pr_dir / "fixt_route_001.pdf").write_bytes(b"%PDF-1.4\n")
    check("...unless the adapter's own exact name is there to settle it",
          _resolve("001") == _pr_dir / "fixt_route_001.pdf")
finally:
    shutil.rmtree(_pr_dir, ignore_errors=True)

print("HF-10: a blank cell's window never encloses another column's glyphs")
import evidence_highway_sequence as _ehs5

_BOUNDS = {"county_city": 100, "city_prefix": 150, "prefix_pm": 200,
           "pm_suffix": 250, "suffix_hg": 300, "hg_ft": 350, "ft_dist": 400,
           "dist_desc": 450}


def _w(x0, x1, text):
    return {"x0": x0, "x1": x1, "top": 100, "bottom": 110, "text": text}


# 'prefix' and 'suffix' print nothing on this row; every other column does.
_PRINTED = {"county": [_w(20, 90, "MON")], "city": [_w(110, 140, "SAL")],
            "pm": [_w(210, 240, "28.013")], "hg": [_w(310, 340, "H")],
            "ft": [_w(360, 390, "U")], "dist": [_w(410, 440, "1.204")]}
_hsl_rec = {"page": 1, "top": 100, "bottom": 110, "desc": [],
            "boundaries": dict(_BOUNDS), "cols": dict(_PRINTED),
            "vals": {"prefix": "", "suffix": ""}}
_others = [word for ws in _PRINTED.values() for word in ws]

_blank_boxes = {}
for _field in ("(col C)", "(col E)"):
    _box = _ehs5.env_box(_hsl_rec, _field)
    _blank_boxes[_field] = _box
    _rect = _box[1] if _box else None
    _hits = [word["text"] for word in _others
             if _rect and word["x0"] < _rect[2] and word["x1"] > _rect[0]]
    check(f"a blank {_field} boxes its own window and no other column's "
          f"glyphs", _box is not None and not _hits)

check("...and the window stays strictly between its own header boundaries",
      _blank_boxes["(col C)"][1][0] >= _BOUNDS["city_prefix"]
      and _blank_boxes["(col C)"][1][2] <= _BOUNDS["prefix_pm"]
      and _blank_boxes["(col E)"][1][0] >= _BOUNDS["pm_suffix"]
      and _blank_boxes["(col E)"][1][2] <= _BOUNDS["suffix_hg"])
check("a printed column still boxes its OWN words, not the window",
      _ehs5.env_box(_hsl_rec, "HG")[1][0] < 310 + 1
      and _ehs5.env_box(_hsl_rec, "HG")[1][2] > 340 - 1)
check("a blank Description REFUSES — the print anchors no cell for it",
      _ehs5.env_box(_hsl_rec, "Description") is None)

# --------------------------------------------------------------------------- #
# HF-10: the env candidate pool comes from the published comparison itself,
# and the aggregate (route-keyed) shape locates by the key when route is blank.
print("HF-10: env candidates derive from the published universe")


class _PubRow:
    def __init__(self, excel_row, route, key, occ, mask, values, token):
        self.excel_row, self.route, self.key = excel_row, route, key
        self.occurrence, self.mask, self.values, self.token = occ, mask, values, token
        self.status = "Both"

    @property
    def matched(self):
        return True

    def state(self, i):
        return self.mask[i]

    def value(self, i):
        return self.values[i]


class _Pub:
    fields = ("F", "G")
    side_labels = ("A", "B")

    def __init__(self, rows):
        self.rows = rows

    def position_of(self, f):
        return self.fields.index(f)

    def is_solo(self, _row):
        return True


class _Ledger5:
    def for_field(self, _f):
        return None


_pub = _Pub([_PubRow(7, "", "001", 1, "DE", ("1 ≠ 2", "9"), "tok1"),
             _PubRow(8, "003", "K2", 1, "ED", ("5", "7 ≠ 8"), "tok2")])
_cand, _misses = ve._env_candidates(_pub, ["F", "G"], _Ledger5())
check("solo differing rows become candidates carrying the published identity; "
      "a route-keyed row (no route column — Ramp Summary) carries its key as "
      "the print-resolution route",
      [e["route"] for e in _cand["F"]] == ["001"]
      and _cand["F"][0]["key"] == "001"
      and _cand["F"][0]["published_token"] == "tok1"
      and _cand["G"][0]["route"] == "003"
      and _cand["G"][0]["display"] == "7 ≠ 8")

# env_fields pins: each adapter's list is DERIVED from the env comparison's own
# constants, so the two can never drift silently.
print("HF-10: env_fields match the env comparison's own display columns")
import compare_env as _ce5
import evidence_ramp_detail as _erd5
import evidence_intersection_detail as _eid5
import evidence_ramp_summary as _ers5
import compare_intersection_detail_tsn as _idt5

check("RD env fields = the forced classic display header + print-only, minus PM",
      _erd5.env_fields()
      == [f for f in (list(_ce5._RD_ENV_HEADER) + ["On/Off", "Ramp Type"])
          if f != "PM"])
check("ID env fields = the canonical export header minus Route/Post Mile",
      _eid5.env_fields()
      == [f for f in _idt5._TSMIS_HEADER if f not in ("Route", "Post Mile")])
import highway_log_columns as _hlc5
check("HL env fields = the corrected column SoT minus the key "
      "(cross-module pin, not the module's own alias — RB-4 audit)",
      _ehl5.env_fields()
      == [f for f in _hlc5.HEADER if f != _ehl5.KEY_LABEL])
check("HSL env fields = the per-route layout with '(col C)'/'(col E)' for the "
      "unnamed postmile columns (compare_env's own relabel rule)",
      _ehs5.env_fields() == ["County", "City", "(col C)", "(col E)", "HG", "FT",
                             "Distance To Next Point", "Description"])
check("RS env fields = RS_HEADER minus Route",
      _ers5.env_fields() == list(_ce5.RS_HEADER[1:]))

# The RS geometry twin: single-line rows parse like the consolidator's own
# two-column walk, and a count/label pair maps to the count word's box.
print("HF-10: the Ramp Summary geometry twin")
import consolidate_ramp_summary as _rs5

_words = [
    {"text": "228", "x0": 20.0, "x1": 40.0, "top": 100.0, "bottom": 108.0},
    {"text": "D", "x0": 50.0, "x1": 56.0, "top": 100.0, "bottom": 108.0},
    {"text": "-", "x0": 58.0, "x1": 62.0, "top": 100.0, "bottom": 108.0},
    {"text": "Divided", "x0": 64.0, "x1": 96.0, "top": 100.0, "bottom": 108.0},
    {"text": "17", "x0": 20.0, "x1": 32.0, "top": 120.0, "bottom": 128.0},
    {"text": "U", "x0": 50.0, "x1": 56.0, "top": 120.0, "bottom": 128.0},
    {"text": "-", "x0": 58.0, "x1": 62.0, "top": 120.0, "bottom": 128.0},
    {"text": "Undivided", "x0": 64.0, "x1": 104.0, "top": 120.0, "bottom": 128.0},
]
_rows5 = _ers5._line_rows_with_boxes(_words, left=True)
check("the twin parses (number, cleaned label, count word) per single line",
      [(n, l) for n, l, _w, _lw in _rows5]
      == [(228, "D - Divided"), (17, "U - Undivided")])
_cells5 = {}
_ers5._attribute(_rows5, _rs5.HIGHWAY_GROUPS, set(), _cells5)
check("attribution finds the schema categories at their count words",
      "hwy_divided" in _cells5 and _cells5["hwy_divided"][2] == 228)
_rec5 = {"record": {"hwy_divided": 228}, "cells": _cells5, "src": "r.pdf",
         "page": 2}
_dv_box = _ers5.env_box(_rec5, "Divided")
check("env_box boxes the COUNT word EXACTLY — its own extent, never the "
      "label's (RB-4 audit: '>= 18' was satisfied by every word on the line)",
      _dv_box is not None
      and _dv_box[1] == (18.0, 98.0, 42.0, 110.0)
      and _dv_box[1][2] < 50.0
      and _ers5.env_value(_rec5, "Divided") == "228")
_rec5b = {"record": {"hwy_divided": 999}, "cells": _cells5, "src": "r.pdf",
          "page": 2}
check("a geometry/parser disagreement refuses the box (never mislabels)",
      _ers5.env_box(_rec5b, "Divided") is None)
check("an absent category has no line and honestly refuses geometry",
      _ers5.env_box(_rec5, "Left") is None)

print()
if _fail:
    print(f"FAILED: {len(_fail)} check(s):")
    for f in _fail:
        print(f"  - {f}")
    sys.exit(1)
print("check_visual_evidence: all checks passed")
