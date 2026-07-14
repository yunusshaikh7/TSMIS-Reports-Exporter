#!/usr/bin/env python3
"""Build a product-consumable raw-TSN development twin for Highway Sequence.

This is deliberately *not* an acceptance artifact: it consumes the frozen
Stage-8 development row cache instead of reparsing the authoritative PDFs.  It
exists only to exercise the product comparison path against every raw TSN
record, including the 46 pre-county EQUATES TO records that the accepted
normalizer necessarily cannot place in its eight-column output.

No product parser, comparator, schema, or evidence module is imported.
"""

from __future__ import annotations

from collections import Counter
import hashlib
import json
from pathlib import Path
from typing import Iterable, Mapping, Sequence
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell


VISUAL_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd"
)
CACHE = VISUAL_ROOT / "phase8_highway_sequence_tsn_rows_draft_r1.json"
STAGE6_RESULT = (
    VISUAL_ROOT / "phase6_tsn_conservation"
    / "highway_sequence_conservation_r7.json"
)
STAGE6_ACCEPTANCE = Path(str(STAGE6_RESULT) + ".acceptance.json")
NORMALIZED = (
    VISUAL_ROOT / "phase4_tsn_rebaseline" / "raw-2026-07-12-r7"
    / "highway_sequence" / "consolidated"
    / "tsn_highway_sequence_normalized.xlsx"
)
OUTPUT_ROOT = VISUAL_ROOT / "phase8_highway_sequence_raw_tsn_twin_dev_r2"

WORKBOOK_NAME = "highway_sequence_raw_tsn_audit_twin.xlsx"
PROVENANCE_NAME = "highway_sequence_raw_tsn_audit_twin.provenance.json"
MANIFEST_NAME = "manifest.json"
RESULT_NAME = "result.json"

SHEET_NAME = "Highway Locations (TSN)"
HEADERS = (
    "Route", "County", "PM", "City", "HG", "FT",
    "Distance To Next Point", "Description",
)

EXPECTED_INPUTS = {
    "development_row_cache": {
        "path": CACHE,
        "bytes": 28_829_216,
        "sha256": "b18d2e077b79920cb1f687f06f8193b25e1d8cd2ebeb1d071b84c22b372598a7",
    },
    "stage6_result": {
        "path": STAGE6_RESULT,
        "bytes": 1_276_684,
        "sha256": "bdd344258ced0e138196c518be2d49ee058f5f9c0f52dea860c328fc3216d1e2",
    },
    "stage6_acceptance": {
        "path": STAGE6_ACCEPTANCE,
        "bytes": 5_934,
        "sha256": "71fe59a5f4676d3b935bcbea380374b14fdccfd77b674ea88148fa18760ffde2",
    },
    "normalized": {
        "path": NORMALIZED,
        "bytes": 2_536_901,
        "sha256": "9dc84c661a9284131baf928767e210a6d708c0a338819fca2b69b907f85dd041",
    },
}

EXPECTED_COUNTS = {
    "raw_records": 69_804,
    "data_records": 68_806,
    "equate_records": 998,
    "pre_county_equates": 46,
    "projectable_records": 69_758,
    "pointer_P": 283,
    "pointer_arrow": 282,
    "pointer_total": 565,
    "projection_description_deltas": 1,
}


class TwinError(RuntimeError):
    pass


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _identity(path: Path) -> dict[str, object]:
    stat = path.stat()
    return {
        "canonical_path": str(path.resolve()),
        "bytes": stat.st_size,
        "sha256": _sha_file(path),
        "mtime_ns": stat.st_mtime_ns,
    }


def _assert_identity(label: str) -> dict[str, object]:
    spec = EXPECTED_INPUTS[label]
    path = Path(spec["path"])
    observed = _identity(path)
    if observed["bytes"] != spec["bytes"] or observed["sha256"] != spec["sha256"]:
        raise TwinError(f"{label} identity drift: {observed}")
    return observed


def _json_bytes(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode("utf-8")


def _write_new(path: Path, payload: bytes) -> None:
    with path.open("xb") as handle:
        handle.write(payload)


def _row_values(record: Mapping[str, object]) -> tuple[object, ...]:
    return (
        record["route"], record["county"], record["pm"], record["city"],
        record["hg"], record["ft"], record["distance"],
        record["description"],
    )


def _ordered_digest(rows: Iterable[Sequence[object]]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update(
            json.dumps(
                list(row), ensure_ascii=False, separators=(",", ":"),
            ).encode("utf-8")
        )
        digest.update(b"\n")
    return digest.hexdigest()


def _validate_embedded_bindings(cache: Mapping[str, object]) -> dict[str, object]:
    bindings = cache.get("bindings")
    if not isinstance(bindings, Mapping):
        raise TwinError("development cache has no bindings object")
    result: dict[str, object] = {}
    for label in ("stage6_result", "stage6_acceptance", "normalized"):
        embedded = bindings.get(label)
        if not isinstance(embedded, Mapping):
            raise TwinError(f"development cache has no embedded {label} binding")
        expected = EXPECTED_INPUTS[label]
        observed_path = Path(str(embedded.get("canonical_path", ""))).resolve()
        expected_path = Path(expected["path"]).resolve()
        exact = (
            observed_path == expected_path
            and embedded.get("size") == expected["bytes"]
            and embedded.get("sha256") == expected["sha256"]
        )
        if not exact:
            raise TwinError(f"embedded {label} binding drift: {embedded}")
        result[label] = {
            "canonical_path": str(observed_path),
            "bytes": embedded["size"],
            "sha256": embedded["sha256"],
            "exact": True,
        }
    return result


def _validate_cache(
    cache: Mapping[str, object],
) -> tuple[list[Mapping[str, object]], list[tuple[object, ...]], dict[str, object]]:
    if cache.get("not_an_acceptance_artifact") is not True:
        raise TwinError("development cache lost its non-acceptance marker")
    if cache.get("audit") != "Highway Sequence TSN development row cache":
        raise TwinError(f"unexpected cache audit label: {cache.get('audit')!r}")

    raw = cache.get("raw_records")
    if not isinstance(raw, list) or not all(isinstance(row, Mapping) for row in raw):
        raise TwinError("raw_records is not a list of objects")
    if len(raw) != EXPECTED_COUNTS["raw_records"]:
        raise TwinError(f"raw record count drift: {len(raw)}")
    rows = [_row_values(record) for record in raw]
    if any(
        value is not None and not isinstance(value, str)
        for row in rows for value in row
    ):
        raise TwinError("raw twin schema contains a non-string, non-blank value")

    kinds = Counter(str(record.get("kind")) for record in raw)
    pre_county = [
        record for record in raw
        if record.get("county") is None
    ]
    if kinds != Counter({"data": 68_806, "equate": 998}):
        raise TwinError(f"record-kind count drift: {kinds}")
    if len(pre_county) != 46 or any(row.get("kind") != "equate" for row in pre_county):
        raise TwinError("the 46 blank-County records are not exactly pre-county equates")

    pointers = Counter(record.get("distance") for record in raw)
    if pointers["*P*"] != 283 or pointers["-------->"] != 282:
        raise TwinError(f"raw pointer-token count drift: {pointers}")

    normalized = cache.get("normalized")
    if not isinstance(normalized, Mapping):
        raise TwinError("development cache has no normalized projection")
    if tuple(normalized.get("headers", ())) != HEADERS:
        raise TwinError(f"normalized schema drift: {normalized.get('headers')}")
    normalized_items = normalized.get("rows")
    if not isinstance(normalized_items, list) or len(normalized_items) != 69_758:
        raise TwinError("normalized projection row count drift")
    normalized_rows = [tuple(item["values"]) for item in normalized_items]
    projected = [
        _row_values(record) for record in raw
        if record.get("county") is not None
    ]
    if len(projected) != 69_758:
        raise TwinError(f"projectable record count drift: {len(projected)}")

    deltas: list[dict[str, object]] = []
    for ordinal, (source, target) in enumerate(zip(projected, normalized_rows, strict=True)):
        for column, (raw_value, normalized_value) in enumerate(
            zip(source, target, strict=True)
        ):
            if raw_value != normalized_value:
                deltas.append({
                    "projected_ordinal": ordinal,
                    "normalized_source_row": normalized_items[ordinal]["source_row"],
                    "field": HEADERS[column],
                    "raw": raw_value,
                    "normalized": normalized_value,
                })
    by_field = Counter(item["field"] for item in deltas)
    expected_deltas = Counter({"Distance To Next Point": 565, "Description": 1})
    if by_field != expected_deltas:
        raise TwinError(f"raw-to-normalized delta drift: {by_field}")
    distance_deltas = [
        item for item in deltas if item["field"] == "Distance To Next Point"
    ]
    if (
        Counter(item["raw"] for item in distance_deltas)
        != Counter({"*P*": 283, "-------->": 282})
        or any(item["normalized"] is not None for item in distance_deltas)
    ):
        raise TwinError("pointer-token projection behavior drift")
    description_delta = [item for item in deltas if item["field"] == "Description"]
    expected_description = {
        "raw": "KEMWATER CHEMICAL PLANT - RT/FRONTAGE ROAD - LT.",
        "normalized": "KEMWATER CHEMICAL PLANT - RT/FRONTAGE, ROAD - LT.",
    }
    if len(description_delta) != 1 or any(
        description_delta[0][key] != value
        for key, value in expected_description.items()
    ):
        raise TwinError(f"raw Description punctuation delta drift: {description_delta}")

    validation = {
        "counts": {
            "raw_records": len(raw),
            "data_records": kinds["data"],
            "equate_records": kinds["equate"],
            "pre_county_equates": len(pre_county),
            "projectable_records": len(projected),
            "pointer_P": pointers["*P*"],
            "pointer_arrow": pointers["-------->"],
            "pointer_total": pointers["*P*"] + pointers["-------->"],
        },
        "raw_records_ordered_sha256": _ordered_digest(
            tuple(record[key] for key in sorted(record)) for record in raw
        ),
        "twin_rows_ordered_sha256": _ordered_digest(rows),
        "raw_to_normalized": {
            "projected_rows": len(projected),
            "normalized_rows": len(normalized_rows),
            "typed_cell_delta_count": len(deltas),
            "typed_cell_deltas_by_field": dict(sorted(by_field.items())),
            "pointer_delta_raw_tokens": dict(
                sorted(Counter(item["raw"] for item in distance_deltas).items())
            ),
            "description_punctuation_delta": description_delta[0],
            "raw_projected_ordered_sha256": _ordered_digest(projected),
            "normalized_ordered_sha256": _ordered_digest(normalized_rows),
        },
    }
    return raw, rows, validation


def _text_cell(sheet, value: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    # Force source text to remain literal, including formula-leading strings.
    cell.data_type = "s"
    return cell


def _write_workbook(path: Path, rows: Sequence[Sequence[object]]) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(SHEET_NAME)
    sheet.append([_text_cell(sheet, header) for header in HEADERS])
    for row in rows:
        sheet.append([
            None if value is None else _text_cell(sheet, str(value))
            for value in row
        ])
    with path.open("xb") as handle:
        workbook.save(handle)
    workbook.close()


def _reopen_workbook(path: Path, expected_rows: Sequence[Sequence[object]]) -> dict[str, object]:
    required_parts = {
        "[Content_Types].xml", "_rels/.rels", "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels", "xl/worksheets/sheet1.xml",
    }
    with ZipFile(path, "r") as package:
        names = package.namelist()
        bad_member = package.testzip()
        duplicate_members = sorted(
            name for name, count in Counter(names).items() if count > 1
        )
        missing_parts = sorted(required_parts - set(names))
    if bad_member is not None or duplicate_members or missing_parts:
        raise TwinError(
            "invalid XLSX package: "
            f"bad={bad_member!r}, duplicates={duplicate_members}, missing={missing_parts}"
        )

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            raise TwinError(f"workbook sheet drift: {workbook.sheetnames}")
        sheet = workbook[SHEET_NAME]
        # A write-only openpyxl sheet may correctly omit optional worksheet
        # dimension metadata.  Verify physical rows/cells by streaming instead
        # of trusting max_row, max_column, or calculate_dimension().  A short
        # tuple can only omit trailing blanks and is padded on the right; any
        # physical cell beyond column H or wholly blank physical row is fatal.
        dimension_metadata = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
        }
        streamed: list[tuple[object, ...]] = []
        padded_trailing_blank_cells = 0
        max_physical_width = 0
        for physical_row, physical_values in enumerate(
            sheet.iter_rows(values_only=True), start=1
        ):
            values = tuple(physical_values)
            max_physical_width = max(max_physical_width, len(values))
            if len(values) > len(HEADERS):
                raise TwinError(
                    f"extra physical cells at workbook row {physical_row}: "
                    f"width {len(values)} > {len(HEADERS)}"
                )
            if not any(value is not None for value in values):
                raise TwinError(f"blank physical workbook row: {physical_row}")
            if len(values) < len(HEADERS):
                padded_trailing_blank_cells += len(HEADERS) - len(values)
                values += (None,) * (len(HEADERS) - len(values))
            streamed.append(values)
        if len(streamed) != 69_805:
            raise TwinError(
                f"physical workbook row count drift: {len(streamed)} != 69,805"
            )
        reopened_headers = streamed[0]
        if reopened_headers != HEADERS:
            raise TwinError(f"reopened header drift: {reopened_headers}")

        reopened = streamed[1:]
        mismatch_samples: list[dict[str, object]] = []
        for ordinal, actual_tuple in enumerate(reopened):
            if ordinal < len(expected_rows) and actual_tuple != tuple(expected_rows[ordinal]):
                if len(mismatch_samples) < 10:
                    mismatch_samples.append({
                        "ordinal": ordinal,
                        "expected": list(expected_rows[ordinal]),
                        "actual": list(actual_tuple),
                    })
        if len(reopened) != len(expected_rows) or mismatch_samples:
            raise TwinError(
                f"workbook reopen mismatch: rows={len(reopened)}/{len(expected_rows)}, "
                f"samples={mismatch_samples}"
            )

        blank_county = sum(row[1] is None for row in reopened)
        pointer_counts = Counter(row[6] for row in reopened)
        if blank_county != 46:
            raise TwinError(f"reopened blank-County count drift: {blank_county}")
        if pointer_counts["*P*"] != 283 or pointer_counts["-------->"] != 282:
            raise TwinError(f"reopened pointer-token count drift: {pointer_counts}")
        reopened_digest = _ordered_digest(reopened)
        expected_digest = _ordered_digest(expected_rows)
        if reopened_digest != expected_digest:
            raise TwinError("reopened ordered row digest drift")
        return {
            "zip_test_passed": True,
            "package_member_count": len(names),
            "duplicate_package_members": duplicate_members,
            "required_parts_present": True,
            "sheet_names_exact": True,
            "sheet_name": SHEET_NAME,
            "optional_dimension_metadata_not_trusted": dimension_metadata,
            "streamed_physical_rows": len(streamed),
            "streamed_logical_columns": len(HEADERS),
            "maximum_streamed_physical_width": max_physical_width,
            "padded_omitted_trailing_blank_cells": padded_trailing_blank_cells,
            "extra_physical_cells": 0,
            "blank_physical_rows": 0,
            "headers": list(reopened_headers),
            "headers_exact": True,
            "data_rows": len(reopened),
            "columns": len(HEADERS),
            "typed_rows_exact": True,
            "ordered_rows_sha256": reopened_digest,
            "blank_county_rows": blank_county,
            "pointer_P": pointer_counts["*P*"],
            "pointer_arrow": pointer_counts["-------->"],
        }
    finally:
        workbook.close()


def main() -> int:
    if OUTPUT_ROOT.exists():
        raise TwinError(f"refusing to overwrite existing output root: {OUTPUT_ROOT}")

    inputs_before = {label: _assert_identity(label) for label in EXPECTED_INPUTS}
    cache = json.loads(CACHE.read_text(encoding="utf-8"))
    embedded_bindings = _validate_embedded_bindings(cache)
    raw, rows, cache_validation = _validate_cache(cache)

    OUTPUT_ROOT.mkdir(parents=False, exist_ok=False)
    workbook_path = OUTPUT_ROOT / WORKBOOK_NAME
    provenance_path = OUTPUT_ROOT / PROVENANCE_NAME
    manifest_path = OUTPUT_ROOT / MANIFEST_NAME
    result_path = OUTPUT_ROOT / RESULT_NAME

    _write_workbook(workbook_path, rows)
    workbook_reopen = _reopen_workbook(workbook_path, rows)

    provenance_rows = []
    for ordinal, (record, values) in enumerate(zip(raw, rows, strict=True), start=2):
        provenance_rows.append({
            "workbook_row": ordinal,
            "source_ref": {
                "member": record["member"],
                "physical_page": record["physical_page"],
                "printed_page": record["printed_page"],
                "line": record["line"],
                "top": record["top"],
            },
            "source_context": {
                "district": record["district"],
                "direction": record["direction"],
                "record_kind": record["kind"],
                "raw_text": record["raw_text"],
            },
            "workbook_values": list(values),
        })
    provenance = {
        "audit": "Highway Sequence raw-TSN development twin provenance",
        "not_an_acceptance_artifact": True,
        "reason": "Consumes a frozen development row cache; final acceptance must reparse immutable TSN PDFs.",
        "schema": {"sheet": SHEET_NAME, "headers": list(HEADERS)},
        "development_cache_binding": inputs_before["development_row_cache"],
        "embedded_accepted_bindings": embedded_bindings,
        "raw_members": cache["raw_members"],
        "raw_documents": cache["raw_documents"],
        "row_count": len(provenance_rows),
        "rows": provenance_rows,
    }
    _write_new(provenance_path, _json_bytes(provenance))

    inputs_after = {label: _assert_identity(label) for label in EXPECTED_INPUTS}
    inputs_unchanged = inputs_before == inputs_after
    if not inputs_unchanged:
        raise TwinError("one or more bound inputs changed during twin construction")

    generated = {
        "workbook": _identity(workbook_path),
        "provenance": _identity(provenance_path),
    }
    manifest = {
        "audit": "Highway Sequence raw-TSN development twin manifest",
        "not_an_acceptance_artifact": True,
        "output_root": str(OUTPUT_ROOT.resolve()),
        "inputs_before": inputs_before,
        "inputs_after": inputs_after,
        "inputs_unchanged": inputs_unchanged,
        "embedded_accepted_bindings": embedded_bindings,
        "generated": generated,
        "cache_validation": cache_validation,
        "workbook_reopen": workbook_reopen,
    }
    _write_new(manifest_path, _json_bytes(manifest))

    result = {
        "audit": "Highway Sequence raw-TSN development twin builder",
        "status": "PASS",
        "terminal": True,
        "not_an_acceptance_artifact": True,
        "reason": "Development-only raw twin built from identity-bound row cache.",
        "output_root": str(OUTPUT_ROOT.resolve()),
        "inputs_unchanged": True,
        "invariants": {
            "exact_product_consumable_8_column_schema": True,
            "raw_records_69804": len(rows) == 69_804,
            "data_records_68806": cache_validation["counts"]["data_records"] == 68_806,
            "equate_records_998": cache_validation["counts"]["equate_records"] == 998,
            "blank_county_pre_county_equates_46": cache_validation["counts"]["pre_county_equates"] == 46,
            "raw_pointer_tokens_565": cache_validation["counts"]["pointer_total"] == 565,
            "raw_vs_normalized_description_punctuation_delta_1": cache_validation["raw_to_normalized"]["typed_cell_deltas_by_field"]["Description"] == 1,
            "xlsx_package_reopened": workbook_reopen["zip_test_passed"],
            "xlsx_headers_rows_cells_exact": workbook_reopen["typed_rows_exact"],
            "xlsx_streamed_69805_physical_rows": workbook_reopen["streamed_physical_rows"] == 69_805,
            "xlsx_streamed_8_logical_columns": workbook_reopen["streamed_logical_columns"] == 8,
            "xlsx_no_extra_cells_or_blank_physical_rows": workbook_reopen["extra_physical_cells"] == 0 and workbook_reopen["blank_physical_rows"] == 0,
            "xlsx_optional_dimension_metadata_not_trusted": True,
            "inputs_unchanged": inputs_unchanged,
        },
        "artifacts": {
            **generated,
            "manifest": _identity(manifest_path),
        },
        "counts": cache_validation["counts"],
        "ordered_rows_sha256": workbook_reopen["ordered_rows_sha256"],
        "description_punctuation_delta": cache_validation["raw_to_normalized"]["description_punctuation_delta"],
    }
    if not all(result["invariants"].values()):
        raise TwinError(f"terminal invariants failed: {result['invariants']}")
    _write_new(result_path, _json_bytes(result))

    print(
        "PASS Highway Sequence raw-TSN development twin: "
        f"{len(rows):,} rows; 46 blank-County equates; 565 pointer tokens; "
        f"{OUTPUT_ROOT}"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, ValueError, KeyError, TypeError, TwinError) as exc:
        print(f"FAIL Highway Sequence raw-TSN development twin: {exc}")
        raise SystemExit(1)
