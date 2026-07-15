"""Golden check for the TSMIS-vs-TSN Ramp Summary comparator
(scripts/compare_ramp_summary_tsn.py) — the v0.17.0 AGGREGATE recipe.

Unlike the FLAT comparators, this one compares ONE statewide category-count table
per side (has_route=False; key = category, value = count). The check locks:
  * the CompareSchema wiring (Category/Count header, key_field 0, TSMIS/TSN sides,
    the extra_sheet_writer that appends the familiar layout);
  * the canonical category list (unique keys; the TSN-only P/V "Dummy" ramp types
    and the grand Total are present);
  * the TSMIS loader SUMMING a consolidated workbook's per-route columns to slugs
    strictly (CMP-AUD-021/022: numeric text parses, fractions/booleans refuse,
    duplicate columns refuse, an absent column stays absent — never a fabricated 0);
  * the independent per-side validation (CMP-AUD-020): every block must partition
    the grand total per the censused contract (the TSMIS Ramp-Types block is
    bounded by the P/V-not-tabulated residual, EXPOSED as a familiar-sheet note),
    all-zero categories under a non-zero total refuse, and a missing category or
    total is a hard stop;
  * end-to-end through compare()/the VALUES workbook (read back with openpyxl, no
    Excel, CI-safe): P/V stay one-sided ('Only in TSN', CMP-AUD-024/025 — never a
    fabricated 0-vs-N), the Ramp Points footnote is display-only, and there is NO
    Route column (has_route=False).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_ramp_summary_tsn.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_ramp_summary_tsn as cmp
import summary_layout
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []
DIFF = " ≠ "          # the diff marker count_diffs / the workbook key on

# slug -> compare key, for building synthetic TSN rows from canonical categories.
_KEY = {slug: key for key, slug in cmp._CATEGORIES}


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _write_tsmis(path, displays, perroute):
    """A synthetic CONSOLIDATED Ramp Summary workbook: row1 group band (ignored),
    row2 the display headers, then one row per route."""
    wb = Workbook()
    ws = wb.active
    ws.title = cmp.TSMIS_SHEET
    ws.append(["group"] * len(displays))
    ws.append(displays)
    for d in perroute:
        ws.append([d.get(x, "") for x in displays])
    wb.save(path)
    wb.close()


# The full consolidated column set (every compared category + total + footnote),
# in the consolidator's own display spelling.
_ALL_DISPLAYS = ["Route"] + list(cmp._SLUG_TO_DISPLAY.values())
_D = cmp._SLUG_TO_DISPLAY   # slug -> display header


def _route_row(route, total, *, right, divided=0, on=None, nolw, diamond):
    """One arithmetically CONSISTENT per-route row (the censused partition
    contract): hwy == total; population == total; on/off + no-linework == total;
    ramp-types + no-linework <= total (the P/V residual). P/V stay blank, like
    every real TSMIS export."""
    d = {disp: 0 for disp in _ALL_DISPLAYS}
    d["Route"] = route
    d[_D["hwy_right"]] = right
    d[_D["hwy_divided"]] = divided
    d[_D["pop_rural_inside"]] = total
    d[_D["onoff_on"]] = (total - nolw) if on is None else on
    d[_D["ramp_points_no_linework"]] = nolw
    d[_D["ramp_D_diamond"]] = diamond
    d[_D["ramp_P_dummy_paired"]] = ""
    d[_D["ramp_V_dummy_volume"]] = ""
    d[_D["total_ramps"]] = total
    return d


def _tsn_rows(total, *, right, divided, undivided, p, v):
    """A CONSISTENT full 31-row normalized TSN table: all four blocks partition
    `total` exactly (P/V included, the censused TSN contract)."""
    vals = {slug: 0 for _k, slug in cmp._CATEGORIES}
    vals["hwy_right"], vals["hwy_divided"], vals["hwy_undivided"] = right, divided, undivided
    vals["onoff_on"] = total
    vals["pop_urban_inside"] = total
    vals["ramp_P_dummy_paired"], vals["ramp_V_dummy_volume"] = p, v
    vals["ramp_D_diamond"] = total - p - v
    vals["total_ramps"] = total
    return [(_KEY[slug], val) for slug, val in vals.items()]


def _write_tsn_norm(path, rows):
    """A synthetic NORMALIZED TSN workbook (Category | Count), keyed on compare keys."""
    wb = Workbook()
    ws = wb.active
    ws.title = cmp.NORMALIZED_SHEET
    ws.append(["Category", "Count"])
    for k, v in rows:
        ws.append([k, v])
    wb.save(path)
    wb.close()


def _sheet(path, name):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
        return header, rows
    finally:
        wb.close()


def test_schema_and_categories():
    print("schema + canonical categories:")
    sc = cmp._SCHEMA
    check("header is Category / Count", sc.header == ["Category", "Count"])
    check("key_field is the category (0)", sc.key_field == 0)
    check("side names TSMIS / TSN", sc.side_a == "TSMIS" and sc.side_b == "TSN")
    check("extra_sheet_writer is set (familiar layout)", sc.extra_sheet_writer is not None)
    keys = [k for k, _s in cmp._CATEGORIES]
    check("category keys are unique", len(keys) == len(set(keys)))
    check("31 compared categories", len(cmp._CATEGORIES) == 31)
    slugs = {s for _k, s in cmp._CATEGORIES}
    check("TSN-only P and V ramp types are in the canonical set",
          {"ramp_P_dummy_paired", "ramp_V_dummy_volume"} <= slugs)
    check("grand Total is a compared category", "total_ramps" in slugs)
    check("Ramp Points w/out linework is a footnote (not a compared category)",
          "ramp_points_no_linework" not in slugs
          and any(f.slug == "ramp_points_no_linework" for f in summary_layout.RAMP_SUMMARY_SPEC.footnotes))


def test_tsmis_loader_sums():
    print("TSMIS loader sums per-route columns to slugs (strict, CMP-AUD-021/022):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_rs_sum_"))
    p = root / "consol.xlsx"
    # Two routes; include P (blank) but OMIT V entirely.
    displays = ["Route", "Right", "Divided", "P-DummyPair", "Total Ramps", "Pts w/o Linework"]
    _write_tsmis(p, displays, [
        {"Route": "001", "Right": 4, "Divided": 10, "P-DummyPair": "", "Total Ramps": 14, "Pts w/o Linework": 2},
        {"Route": "002", "Right": "6", "Divided": 10, "P-DummyPair": "", "Total Ramps": 16, "Pts w/o Linework": 1},
    ])
    s = cmp._load_tsmis(p)
    check("hwy_right summed, numeric TEXT parsed (4+'6'=10, not dropped)",
          s["hwy_right"] == 10)
    check("hwy_divided summed (10+10=20)", s["hwy_divided"] == 20)
    check("blank P column is PRESENT with 0 (column exists, cells blank)",
          s["ramp_P_dummy_paired"] == 0)
    check("missing V column stays ABSENT (never a fabricated 0)",
          "ramp_V_dummy_volume" not in s)
    check("total_ramps summed (14+16=30)", s["total_ramps"] == 30)
    check("ramp_points summed (2+1=3)", s["ramp_points_no_linework"] == 3)

    # Strict count refusals, with the file + column named (CMP-AUD-021).
    for label, bad, needle in (("fractional 1.9", 1.9, "fractional"),
                               ("boolean True", True, "boolean"),
                               ("malformed text 'ten'", "ten", "not a whole number")):
        pb = root / f"consol_{needle.split()[0]}.xlsx"
        _write_tsmis(pb, displays, [
            {"Route": "001", "Right": bad, "Divided": 10, "P-DummyPair": "",
             "Total Ramps": 14, "Pts w/o Linework": 2}])
        try:
            cmp._load_tsmis(pb)
            check(f"{label} count refuses", False)
        except ValueError as e:
            check(f"{label} count refuses naming file + column",
                  needle in str(e) and pb.name in str(e) and "'Right'" in str(e))

    # A duplicated category column refuses (CMP-AUD-022).
    pd = root / "consol_dupcol.xlsx"
    _write_tsmis(pd, displays + ["Right"], [
        {"Route": "001", "Right": 4, "Divided": 10, "P-DummyPair": "",
         "Total Ramps": 14, "Pts w/o Linework": 2}])
    try:
        cmp._load_tsmis(pd)
        check("duplicated category column refuses", False)
    except ValueError as e:
        check("duplicated category column refuses", "duplicated" in str(e))


def test_end_to_end():
    print("end-to-end VALUES workbook (aggregate compare + familiar sheet):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_rs_tsn_"))
    tsmis_path = root / "tsmis.xlsx"
    tsn_path = root / "tsn.xlsx"
    out_path = root / "cmp.xlsx"
    # Two CONSISTENT routes (total 30): hwy right 10 + divided 20; on 27 + nolw 3;
    # rural-inside 30; diamond 25 + nolw 3 = 28 -> P/V residual 2 (exposed note).
    _write_tsmis(tsmis_path, _ALL_DISPLAYS, [
        _route_row("001", 14, right=4, divided=10, nolw=2, diamond=11),
        _route_row("002", 16, right=6, divided=10, nolw=1, diamond=14),
    ])
    # CONSISTENT full TSN table (total 40): right matches (10), divided differs
    # (25), undivided differs (0 vs 5), P/V carried (5/3), diamond 32, total 40.
    _write_tsn_norm(tsn_path, _tsn_rows(40, right=10, divided=25, undivided=5, p=5, v=3))
    res = cmp.compare(tsmis_path, tsn_path, out_path, events=Events(),
                      confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")

    header, rows = _sheet(out_path, "Comparison")
    check("NO Route column (has_route=False)", "Route" not in header)
    cat_col = header.index("Category")
    cnt_col = header.index("Count")
    status_col = header.index("Status")
    by_cat = {r[cat_col]: r for r in rows}

    both = sum(1 for r in rows if r[status_col] == "Both")
    tsmis_only = sum(1 for r in rows if r[status_col] == "TSMIS only")
    tsn_only = sum(1 for r in rows if r[status_col] == "TSN only")
    # CMP-AUD-024/025: P and V are TSN-only (not fabricated TSMIS zeros); the linework
    # footnote is display-only, so there are NO TSMIS-only comparison rows.
    check("29 shared categories matched on both sides", both == 29)
    check("no TSMIS-only comparison rows (footnote is display-only)", tsmis_only == 0)
    check("P and V are the two TSN-only rows", tsn_only == 2)

    ndiff = sum(1 for r in rows if DIFF in r[cnt_col])
    check("exactly 7 differing SHARED categories "
          "(Divided/Undivided/On/Rural-I/Urban-I/Diamond/Total)", ndiff == 7)
    check("P - Dummy Paired is Only in TSN (value 5), not a fabricated 0-vs-5 diff",
          by_cat[_KEY["ramp_P_dummy_paired"]][status_col] == "TSN only")
    check("matching category (Right=10) shows no diff marker",
          DIFF not in by_cat[_KEY["hwy_right"]][cnt_col])

    # Familiar layout sheet present + readable.
    fh, fr = _sheet(out_path, summary_layout.RAMP_SUMMARY_SPEC.sheet_name)
    flat = [c for row in [fh] + fr for c in row]
    check("familiar sheet labels sides TSMIS and TSN", "TSMIS" in flat and "TSN" in flat)
    check("familiar sheet lists the P - Dummy Paired row",
          any("P - Dummy Paired" in c for c in flat))
    # CMP-AUD-184: the shared note must describe the cells truthfully — a
    # one-sided category stays BLANK on the absent side, never a claimed 0.
    check("familiar note says one-sided categories stay BLANK (no zero-fill claim)",
          any("stays BLANK" in c and "real source zero" in c for c in flat)
          and not any("show 0 on that side" in c for c in flat))
    p_row = next((r for r in fr if r and r[0] == "P - Dummy Paired"), None)
    check("P row shows BLANK TSMIS / 5 TSN / BLANK Δ on the familiar sheet",
          p_row is not None and p_row[1] == "" and p_row[2] == "5" and p_row[3] == "")
    check("familiar sheet shows the Ramp Points footnote",
          any("Ramp Points w/out linework" in c for c in flat))
    # CMP-AUD-020: the bounded Ramp-Types residual (2 ramps in TSN-only P/V
    # classes) is EXPOSED as a note on the familiar sheet, never fabricated
    # into a category and never a warning.
    check("familiar sheet exposes the P/V residual note (28 of 30, 2 not)",
          any("Ramp Types" in c and "2 not" in c and "TSMIS" in c for c in flat))
    # CMP-AUD-024: the footnote value (2+1=3) reaches the display sheet out of band,
    # keyed by footnote.key, without ever being a compared row.
    check("footnote value rides the out-of-band channel (2+1=3)",
          cmp._footnote_values(cmp._load_tsmis(tsmis_path))
          == {"Ramp Points w/out linework": 3})
    check("the footnote is not a compared category on either side",
          not any("Ramp Points" in str(r[cat_col]) for r in rows))
    print(f"      (both={both}, TSN-only={tsn_only}, TSMIS-only={tsmis_only}, diffs={ndiff})")


def test_validation_refusals():
    """CMP-AUD-020/021/022: a table that does not reconcile refuses with a named
    block; missing categories/totals are hard stops; duplicates refuse; the
    contract-change tripwire fires on a non-zero TSMIS P/V count."""
    print("independent per-side validation (CMP-AUD-020/021/022):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_rs_val_"))
    good_tsn = root / "tsn_good.xlsx"
    _write_tsn_norm(good_tsn, _tsn_rows(40, right=10, divided=25, undivided=5, p=5, v=3))
    good_tsmis = root / "tsmis_good.xlsx"
    _write_tsmis(good_tsmis, _ALL_DISPLAYS,
                 [_route_row("001", 14, right=4, divided=10, nolw=2, diamond=11)])

    def refuses(label, tsmis, tsn, *needles):
        try:
            cmp._load_pair(tsmis, tsn)
            check(label, False)
        except ValueError as e:
            check(label, all(n in str(e) for n in needles))

    # All-zero categories under a non-zero total (both sides agreeing!) refuse.
    z = _route_row("001", 10, right=0, divided=0, nolw=0, diamond=0)
    z[_D["pop_rural_inside"]] = 0
    z[_D["onoff_on"]] = 0
    zp = root / "tsmis_zero.xlsx"
    _write_tsmis(zp, _ALL_DISPLAYS, [z])
    refuses("all-zero TSMIS categories + total=10 refuse naming the block",
            zp, good_tsn, "Highway Groups", "does not reconcile")
    zn = root / "tsn_zero.xlsx"
    _write_tsn_norm(zn, [(k, 10 if s == "total_ramps" else 0)
                         for k, s in cmp._CATEGORIES])
    refuses("all-zero TSN categories + total=10 refuse too",
            good_tsmis, zn, "does not reconcile")

    # A missing required category / total is a hard stop, never a fabricated 0.
    displays_missing = [d for d in _ALL_DISPLAYS if d != _D["hwy_right"]]
    row = _route_row("001", 14, right=4, divided=10, nolw=2, diamond=11)
    row.pop(_D["hwy_right"])
    mp = root / "tsmis_missing.xlsx"
    _write_tsmis(mp, displays_missing, [row])
    refuses("missing TSMIS category column is a hard stop naming the category",
            mp, good_tsn, "R - Right", "missing")
    tn = root / "tsn_missing_total.xlsx"
    _write_tsn_norm(tn, [(k, 0) for k, s in cmp._CATEGORIES if s != "total_ramps"])
    refuses("missing TSN grand total is a hard stop",
            good_tsmis, tn, "grand total")

    # Duplicate exact normalized key refuses (was: last-wins, CMP-AUD-022).
    dn = root / "tsn_dup.xlsx"
    _write_tsn_norm(dn, [(_KEY["hwy_right"], 4), (_KEY["hwy_right"], 7)])
    try:
        cmp._load_tsn(dn)
        check("duplicate exact normalized key refuses (no last-wins)", False)
    except ValueError as e:
        check("duplicate exact normalized key refuses (no last-wins)",
              "twice" in str(e))

    # Fractional normalized count refuses (was: silent truncation, CMP-AUD-021).
    fn = root / "tsn_frac.xlsx"
    _write_tsn_norm(fn, [(_KEY["hwy_right"], 1.9)])
    try:
        cmp._load_tsn(fn)
        check("fractional normalized count refuses", False)
    except ValueError as e:
        check("fractional normalized count refuses", "fractional" in str(e))

    # The P/V contract-change tripwire: a NON-ZERO TSMIS P count means the form
    # started tabulating the TSN-only classes — refuse instead of one-siding it.
    pv = _route_row("001", 14, right=4, divided=10, nolw=2, diamond=11)
    pv[_D["ramp_P_dummy_paired"]] = 3
    pp = root / "tsmis_pv.xlsx"
    _write_tsmis(pp, _ALL_DISPLAYS, [pv])
    refuses("non-zero TSMIS P count trips the contract-change refusal",
            pp, good_tsn, "Dummy Paired", "update")


def test_provenance_sidecar():
    """CMP-AUD-076: a committed comparison persists a durable provenance record —
    the recipe, each input's FULL canonical selection (basenames are ambiguous),
    its pre-read sha256 + stat identity, and the committed generation binding.
    Same basenames in different directories stay distinguishable; a byte-copy
    records the same digest under its own selection; absence reads as None."""
    print("durable comparison provenance (CMP-AUD-076):")
    import shutil

    import compare_tsn_common as ctc
    root = Path(tempfile.mkdtemp(prefix="tsmis_rs_prov_"))
    a_dir, b_dir = root / "A", root / "B"
    a_dir.mkdir(), b_dir.mkdir()
    # SAME basename on both sides, different directories + different bytes.
    tsmis_path = a_dir / "same.xlsx"
    tsn_path = b_dir / "same.xlsx"
    _write_tsmis(tsmis_path, _ALL_DISPLAYS,
                 [_route_row("001", 14, right=4, divided=10, nolw=2, diamond=11)])
    _write_tsn_norm(tsn_path, _tsn_rows(40, right=10, divided=25, undivided=5, p=5, v=3))
    out = root / "cmp.xlsx"
    res = cmp.compare(tsmis_path, tsn_path, out, events=Events(),
                      confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")

    prov = ctc.read_comparison_provenance(out)
    check("a provenance sidecar exists beside the workbook", prov is not None)
    check("schema + recipe recorded (report + banner)",
          prov.get("schema_version") == 1
          and prov.get("recipe", {}).get("report") == "Ramp Summary"
          and "TSMIS vs TSN" in prov.get("recipe", {}).get("banner", ""))
    ins = prov.get("inputs") or []
    check("both inputs recorded with roles", [i.get("role") for i in ins]
          == ["TSMIS", "TSN"])
    check("FULL canonical selections disambiguate the same basenames",
          all(i["name"] == "same.xlsx" for i in ins)
          and ins[0]["selection"] != ins[1]["selection"]
          and str(a_dir) in ins[0]["selection"] and str(b_dir) in ins[1]["selection"])
    check("distinct content digests + stat identity captured",
          ins[0]["sha256"] != ins[1]["sha256"]
          and all(len(i["sha256"]) == 64 and i["size"] > 0 and i["mtime_ns"] > 0
                  for i in ins))
    check("the record binds the committed generation + member digests",
          prov.get("generation_id") == res.artifact_generation.generation_id
          and prov.get("members") == dict(res.artifact_generation.content_digests))

    # The concise human display: a Provenance SHEET inside the workbook carrying
    # the full selections + digests (the sidecar remains the machine binding).
    ph, pr = _sheet(out, "Provenance")
    flatp = " ".join(c for row in [ph] + pr for c in row)
    check("a Provenance sheet shows both full selections + digests",
          str(a_dir) in flatp and str(b_dir) in flatp
          and ins[0]["sha256"] in flatp and ins[1]["sha256"] in flatp
          and "captured before the inputs were read" in flatp)

    # A byte-copy elsewhere records the SAME digest under its OWN selection.
    c_dir = root / "C"
    c_dir.mkdir()
    copied = c_dir / "same.xlsx"
    shutil.copy2(tsmis_path, copied)
    out2 = root / "cmp2.xlsx"
    res2 = cmp.compare(copied, tsn_path, out2, events=Events(),
                       confirm_overwrite=lambda _p: True, mode="values")
    prov2 = ctc.read_comparison_provenance(out2)
    check("a copy keeps the digest but records its own selection",
          res2.status == "ok"
          and prov2["inputs"][0]["sha256"] == ins[0]["sha256"]
          and prov2["inputs"][0]["selection"] != ins[0]["selection"])

    check("absence reads as None (older comparison), never fabricated",
          ctc.read_comparison_provenance(root / "never_built.xlsx") is None)
    junk = root / "junk.xlsx"
    ctc.provenance_path(junk).write_text("{not json", encoding="utf-8")
    check("a corrupt sidecar reads as None (logged), never trusted",
          ctc.read_comparison_provenance(junk) is None)


def test_corrupt_pdf_is_valueerror():
    """A corrupt/truncated statewide PDF must honor the loader contract:
    ValueError (run_files_compare reports it cleanly), never a raw pdfplumber
    exception escaping into the matrix path."""
    import tempfile
    bad = Path(tempfile.mkdtemp()) / "TSN statewide.pdf"
    bad.write_bytes(b"%PDF-1.4 not really a pdf, just junk bytes with no xref")
    try:
        cmp._load_tsn(bad)
        check("corrupt PDF raises", False)
    except ValueError as e:
        check("corrupt PDF -> ValueError (loader contract)", True)
        check("...message names the file", "TSN statewide.pdf" in str(e))
    except Exception as e:  # noqa: BLE001 — the point of the test
        check(f"corrupt PDF -> ValueError, not {type(e).__name__}", False)


def main():
    test_schema_and_categories()
    test_tsmis_loader_sums()
    test_end_to_end()
    test_validation_refusals()
    test_provenance_sidecar()
    test_corrupt_pdf_is_valueerror()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-RAMP-SUMMARY-TSN CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
