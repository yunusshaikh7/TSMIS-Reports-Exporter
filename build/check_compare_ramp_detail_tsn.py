"""Golden check for the TSMIS-vs-TSN Ramp Detail comparator
(scripts/compare_ramp_detail_tsn.py) — the reference v0.17.0 vs-TSN recipe.

Locks: the CompareSchema wiring (PM key + the TSN-only context fields), route
extraction from the TSN LOCATION, PM/date/description normalization (the TSMIS
export-added route prefix is stripped; TSN text is preserved byte-for-byte —
CMP-AUD-135), the position-based TSMIS-consolidated loader, the D4 county-aware
physical key (CMP-AUD-045: route + county + norm_pm — the Comparison sheet key
column shows the canonical "route / county / pm" display; a v3 normalized
library without the District/PM-Suffix columns refuses with a rebuild hint),
District as a compared field (CMP-AUD-185), the key collapsing a mid-list
insert to a single one-sided ramp (no phantom cascade), and — the property the
opt-in `context_fields` exists for — that a context column NEVER contributes a diff
cell while a compared column does. End-to-end through the real compare()/VALUES
workbook, read back with openpyxl (no Excel, CI-safe).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_ramp_detail_tsn.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_ramp_detail_tsn as rd
import compare_tsn_common as ctc
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []
DIFF = " ≠ "          # the ≠ marker count_diffs / the workbook key on


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


# Consolidated TSMIS layout BY POSITION (Route prepended; labels shift right of
# City Code/R/U/Description, which is why the loader reads by position).
_TSMIS_HDR = ["Route", "Location", "", "PM", "Date of Record", "", "HG", "Area 4",
              "", "City Code", "R/U", "Description"]


def _tsmis_row(route, loc, pr, pm, date, hg, area4, city, ru, desc):
    return [route, loc, pr, pm, date, "", hg, area4, city, ru, desc, ""]


def _write_tsmis(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = rd.TSMIS_SHEET
    ws.append(_TSMIS_HDR)
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _write_tsn(path, rows, sidecars=("TSN District", "TSN County", "TSN PM Suffix"),
               marker_version="current"):
    """rows: [route, PR, PM, District, Date, HG, Area4, City, R/U, Desc,
    RampName, OnOff, RampType, ADT, dist, cnty, sfx] — the v4 normalized shape
    (District in the shared width + the District/County/PM-Suffix sidecars).
    Stamps the CMP-AUD-037 normalization marker at the current version by
    default so _load_tsn accepts it; pass marker_version=None to model a
    pre-marker library, or an int to model a stale one."""
    wb = Workbook()
    ws = wb.active
    ws.title = rd.NORMALIZED_SHEET
    ws.append(["Route"] + rd.SHARED_HEADER + list(sidecars))
    for r in rows:
        ws.append(r)
    if marker_version == "current":
        marker_version = rd.NORMALIZATION_VERSION
    if marker_version is not None:
        ctc.write_normalization_marker(wb, marker_version)
    wb.save(path)
    wb.close()


def _comparison(path):
    """(header, rows) of the Comparison sheet from a VALUES workbook."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Comparison"]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
        return header, rows
    finally:
        wb.close()


# The July-2026 consolidated layout (HF-04): every label over its own value, no
# PM-suffix column, OF/TY between R/U and Description. Kept as a literal here
# (not read from the module) so the check still binds the exact censused
# header if the module's copy drifts.
_TSMIS_HDR_2026 = ["Route", "Location", "PRE", "PM", "Date of Record", "HG",
                   "Area 4", "City Code", "R/U", "OF", "TY", "Description"]


def _tsmis_row_2026(route, loc, pre, pm, date, hg, area4, city, ru, of, ty, desc):
    return [route, loc, pre, pm, date, hg, area4, city, ru, of, ty, desc]


def _write_tsmis_2026(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = rd.TSMIS_SHEET
    ws.append(_TSMIS_HDR_2026)
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def test_dual_layout_loader():
    """HF-04 / PCOA-FINAL-001: BOTH censused consolidated layouts load, each by
    its own position map (OF/TY/Description mapped correctly — the exact
    mis-mapping hazard the pinned single layout protected against); an unknown
    third layout refuses with a message that names the real gate instead of
    misdiagnosing a missing leading 'Route' column."""
    print("HF-04 — dual-layout consolidated loader:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_rd_hf04_"))
    new_path = root / "new.xlsx"
    # The finding's witness shape: route 001 row 2 (OF=F TY=D) + a null-token
    # row (the 005 / 025.218 class) + a genuine on-ramp letter.
    _write_tsmis_2026(new_path, [
        _tsmis_row_2026("001", "12-ORA-001", "R", "000.606", "02/25/1976", "D",
                        "Y", "DAPT", "U", "F", "D", "001/NB OFF TO DOHENY PK RD"),
        _tsmis_row_2026("005", "07-LA-005", "", "025.218", "01/01/2014", "D",
                        "-", "LA", "U", "-", "-", "NO RAMP LINEAR EVENT"),
        _tsmis_row_2026("005", "07-LA-005", "R", "026.000", "01/01/2014", "D",
                        "Y", "LA", "U", "N", "C", "005/NB ON FR X"),
    ])
    hdr = ["Route"] + rd.SHARED_HEADER
    try:
        rows, has_route = rd._load_tsmis(new_path)
    except ValueError as e:
        # The pre-fix tree refuses the July-2026 layout here — with the
        # misdiagnosing "expected a leading 'Route' column" message the
        # finding recorded. Degrade to semantic FAILs, never a crash.
        print(f"     -> pre-fix refusal: {str(e)[:110]}")
        for name in ("the July-2026 layout is accepted",
                     "Description read from position 11 (route prefix stripped)",
                     "OF maps to the On/Off context column",
                     "TY maps to the Ramp Type context column",
                     "HG/Area 4/City Code/R/U at the July positions",
                     "the July layout conserves an empty PM-suffix claim",
                     "the null-token row loads byte-exact "
                     "('-' / 'NO RAMP LINEAR EVENT')"):
            check(name, False)
    else:
        check("the July-2026 layout is accepted",
              has_route is True and len(rows) == 3)
        by = {str(r[1 + rd.KEY_FIELD]): r for r in rows}
        r2 = by["0.606"]
        check("Description read from position 11 (route prefix stripped)",
              r2[hdr.index("Description")] == "NB OFF TO DOHENY PK RD")
        check("OF maps to the On/Off context column",
              r2[hdr.index("On/Off")] == "F")
        check("TY maps to the Ramp Type context column",
              r2[hdr.index("Ramp Type")] == "D")
        check("HG/Area 4/City Code/R/U at the July positions",
              (r2[hdr.index("HG")], r2[hdr.index("Area 4")],
               r2[hdr.index("City Code")], r2[hdr.index("R/U")])
              == ("D", "Y", "DAPT", "U"))
        check("the July layout conserves an empty PM-suffix claim",
              dict((c.name, c.value) for c in
                   r2[1 + rd.KEY_FIELD].physical_identity.raw_claims)
              ["postmile_suffix"] == "")
        check("the null-token row loads byte-exact ('-' / 'NO RAMP LINEAR EVENT')",
              (by["25.218"][hdr.index("Area 4")],
               by["25.218"][hdr.index("On/Off")],
               by["25.218"][hdr.index("Description")])
              == ("-", "-", "NO RAMP LINEAR EVENT"))

    # The classic layout still loads (the existing fixtures above prove the
    # projection; this pins acceptance through the SAME dispatching loader).
    old_path = root / "old.xlsx"
    _write_tsmis(old_path, [
        _tsmis_row("001", "12-ORA-001", "R", "000.606", "02/25/1976", "D",
                   "Y", "DAPT", "U", "001/NB OFF TO DOHENY PK RD")])
    rows_old, _ = rd._load_tsmis(old_path)
    check("the classic layout still loads through the same gate",
          len(rows_old) == 1
          and rows_old[0][hdr.index("Description")] == "NB OFF TO DOHENY PK RD"
          and rows_old[0][hdr.index("On/Off")] == "")

    # An UNKNOWN third layout refuses, and the message names the real gate: it
    # must not misdiagnose the leading 'Route' column the workbook has, and it
    # must name the supported-editions bind (the PCOA-FINAL-001 message fix).
    unknown = root / "unknown.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = rd.TSMIS_SHEET
    ws.append(["Route", "Location", "PRE", "PM", "Date of Record", "HG",
               "Area 4", "City Code", "R/U", "OF", "TY", "XS", "Description"])
    ws.append(["001", "12-ORA-001", "R", "1.000", "2026-01-01", "D", "Y", "C",
               "U", "F", "D", "x", "001/DESC"])
    wb.save(unknown)
    wb.close()
    try:
        rd._load_tsmis(unknown)
        check("an unknown layout is refused", False)
    except ValueError as e:
        msg = str(e)
        check("an unknown layout is refused", True)
        check("the refusal names the supported-layout gate",
              "supported layout" in msg and "July-2026" in msg)
        check("the refusal no longer misdiagnoses the leading 'Route' column",
              "expected a leading 'Route' column" not in msg)

    # The consolidator's consumability predicate mirrors the loader's gate.
    ok_pred = getattr(rd, "consolidated_header_ok", None)
    check("consolidated_header_ok exists and accepts both editions",
          ok_pred is not None and ok_pred(rd._TSMIS_HEADER)
          and ok_pred(_TSMIS_HDR_2026))
    check("consolidated_header_ok refuses junk",
          ok_pred is not None
          and not ok_pred(["Route"] + ["x"] * 11)
          and not ok_pred(_TSMIS_HDR_2026[:-1]))


def test_consolidator_completion_truth():
    """HF-04 / PCOA-FINAL-001 defect (a): consolidate 126/126-style green for a
    workbook no comparator accepts must not happen. An unknown-layout input
    folder consolidates into a kept-for-inspection file but the RESULT is an
    error with completion=failed; both accepted editions still report
    ok/complete."""
    print("HF-04 — consolidator completion agrees with consumability:")
    import consolidate_ramp_detail as crd
    import outcome
    from events import Events
    root = Path(tempfile.mkdtemp(prefix="tsmis_rd_hf04c_"))

    def consolidate_from(header, row):
        src = root / f"in_{len(list(root.iterdir()))}"
        src.mkdir()
        wb = Workbook()
        ws = wb.active
        ws.title = rd.TSMIS_SHEET
        ws.append(header)
        ws.append(row)
        wb.save(src / "tsar_ramp_detail_route_001.xlsx")
        wb.close()
        out = src / "consolidated.xlsx"
        res = crd.consolidate(events=Events(),
                              confirm_overwrite=lambda _p: True,
                              input_dir=src, out_path=out)
        return res, out

    # Unknown layout: the export gained a column this app doesn't know.
    res, out = consolidate_from(
        ["Location", "PRE", "PM", "Date of Record", "HG", "Area 4",
         "City Code", "R/U", "OF", "TY", "XS", "Description"],
        ["12-ORA-001", "R", "1.000", "2026-01-01", "D", "Y", "C", "U", "F",
         "D", "x", "001/DESC"])
    check("an unknown-layout consolidation does NOT report ok",
          res.status != "ok")
    check("...and its completion is failed (not comparable, not promotable)",
          getattr(res, "completion", None) == outcome.FAILED)
    check("...the combined file is kept for inspection and named in the message",
          out.exists() and str(out) in (res.message or ""))
    check("...and the message names the supported editions",
          "July-2026" in (res.message or ""))

    # Both accepted editions still consolidate green.
    res_old, _ = consolidate_from(
        list(rd._TSMIS_HEADER[1:]),
        ["12-ORA-001", "R", "1.000", "2026-01-01", "", "D", "Y", "C", "U",
         "001/DESC", ""])
    check("a classic-layout consolidation reports ok/complete",
          res_old.status == "ok"
          and getattr(res_old, "completion", None) == outcome.COMPLETE)
    res_new, _ = consolidate_from(
        _TSMIS_HDR_2026[1:],
        ["12-ORA-001", "R", "1.000", "2026-01-01", "D", "Y", "C", "U", "F",
         "D", "001/DESC"])
    check("a July-2026-layout consolidation reports ok/complete",
          res_new.status == "ok"
          and getattr(res_new, "completion", None) == outcome.COMPLETE)


def test_schema():
    print("schema wiring:")
    sc = rd._SCHEMA
    check("key_field is PM", sc.header[sc.key_field] == "PM" and sc.key_field == rd.KEY_FIELD)
    check("side names TSMIS / TSN", sc.side_a == "TSMIS" and sc.side_b == "TSN")
    check("context_fields = the 4 TSN-only columns",
          set(sc.context_fields) == {"Ramp Name", "On/Off", "Ramp Type", "ADT"})
    check("Date of Record is a date field", "Date of Record" in sc.date_fields)
    raw_probe = rd._tsn_raw_row(
        ["01-DN-101", "R", "001.000", "E", "2026-01-01", "D", "Y", "C", "U", "9/DESC"],
        {"LOCATION": 0, "PR": 1, "PM": 2, "PM_SFX": 3, "DATE_OF_RECORD": 4,
         "HG": 5, "AREA_4": 6, "CITY_CODE": 7, "POP": 8, "DESCRIPTION": 9})
    check("route from TSN LOCATION '01-DN-101' -> '101'", raw_probe[0] == "101")
    check("the PM key carries the D4 identity (route/county/DECIMAL-canonical "
          "PM — CMP-AUD-006: '9.6', '9.600', '009.600' are ONE ramp)",
          dict(raw_probe[1 + rd.KEY_FIELD].physical_identity.canonical_components)
          == {"route": "101", "county": "DN", "postmile": "1"})
    check("District is a compared field filled from LOCATION",
          rd.SHARED_HEADER[2] == "District" and raw_probe[3] == "01")
    check("TSN Description preserved byte-for-byte (its own '9/' prefix survives)",
          raw_probe[1 + rd.SHARED_HEADER.index("Description")] == "9/DESC")
    check("PM normalizes ' 000.606' and '0.606' to the same canon",
          rd._norm_pm(" 000.606") == rd._norm_pm("0.606") == "0.606")
    check("date ISO from both formats",
          rd._iso_date("02/25/1976") == "1976-02-25"
          and rd._iso_date("1992-09-28 00:00:00") == "1992-09-28")
    check("description drops the TSMIS '001/' route prefix",
          rd._strip_desc_prefix("001/NB OFF TO DOHENY") == "NB OFF TO DOHENY")
    # CMP-AUD-197 (the RD vs-TSN half): the loader reads cells the way
    # installed Excel does — OOXML _xHHHH_ escapes decode at the load
    # boundary. The bound raw TSN extract carries ZERO literal _x000d_, so
    # the four Cactus City Excel cells (route 010) were export-encoding
    # artifacts, not data differences.
    check("OOXML escapes decode at the load boundary (both hex cases)",
          rd._v("A_x000d_B") == "A\rB" and rd._v("A_x000D_B") == "A\rB")
    check("the _x005F_ escaped literal underscore is preserved",
          rd._v("TAG_x005F_x000d_") == "TAG_x000d_")
    check("TSMIS Description decodes BEFORE the edge trim (Cactus City)",
          rd._strip_desc_prefix("010/EBOFF TO CACTUS_x000d_\n", "010")
          == "EBOFF TO CACTUS")
    check("TSN-side Description edge-trim decodes consistently",
          rd._edge_text("EBOFF TO CACTUS\n") == "EBOFF TO CACTUS"
          and rd._edge_text("X_x000d_") == "X")
    check("a decoded INTERIOR carriage return survives as real content",
          rd._v("A_x000d_ B") == "A\r B")


def test_end_to_end():
    print("end-to-end VALUES workbook (counts + context non-asserting):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_rd_tsn_"))
    # Side A = the TSMIS consolidated Ramp Detail export; side B = the raw TSN
    # workbook; out_path = the comparison workbook compare() writes.
    tsmis_path = root / "tsmis.xlsx"
    tsn_path = root / "tsn.xlsx"
    out_path = root / "cmp.xlsx"
    # Two routes. Matched PMs with: an identical row, a COMPARED diff (HG), a
    # CONTEXT-only value (Ramp Type, TSMIS blank), and one-sided ramps on each side.
    _write_tsmis(tsmis_path, [
        _tsmis_row("001", "12-ORA-001", "R", "000.606", "02/25/1976", "D", "Y", "DAPT", "U", "001/NB OFF TO X"),
        _tsmis_row("001", "12-ORA-001", "R", "001.000", "02/25/1976", "",  "Y", "DAPT", "U", "001/SB ON FR Y"),   # HG blank vs TSN 'D' -> COMPARED diff
        _tsmis_row("001", "12-ORA-001", "R", "002.000", "01/01/2000", "D", "N", "LGNB", "U", "001/RAMP Z"),       # Ramp Type context only
        _tsmis_row("002", "12-ORA-002", "M", "010.000", "03/03/1990", "U", "Y", "SANA", "R", "002/ON A"),
    ])
    _write_tsn(tsn_path, [
        ["001", "R", "0.606", "12", "1976-02-25", "D", "Y", "DAPT", "U", "NB OFF TO X", "101_1", "F", "D", "70", "12", "ORA", ""],
        ["001", "R", "1.000", "12", "1976-02-25", "D", "Y", "DAPT", "U", "SB ON FR Y", "101_2", "O", "F", "80", "12", "ORA", ""],   # HG 'D' vs TSMIS blank
        ["001", "R", "1.500", "12", "1965-01-01", "D", "Y", "LGNB", "U", "MID INSERT", "101_x", "O", "H", "90", "12", "ORA", ""],   # only in TSN (mid-list)
        ["001", "R", "2.000", "12", "2000-01-01", "D", "N", "LGNB", "U", "RAMP Z", "101_3", "F", "M", "55", "12", "ORA", ""],       # Ramp Type 'M' (TSMIS blank) -> context, no diff
        ["002", "M", "10.000", "12", "1990-03-03", "U", "Y", "SANA", "R", "ON A", "201_1", "O", "D", "30", "12", "ORA", ""],
    ])
    res = rd.compare(tsmis_path, tsn_path, out_path, events=Events(), confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows = _comparison(out_path)

    # Key collapse: the mid-list TSN insert is ONE one-sided ramp, no cascade.
    # The Comparison key column shows the side-independent canonical identity
    # display "route / county / postmile" (CMP-AUD-045).
    pm_col = header.index("PM")
    by_pm = {r[pm_col]: r for r in rows}
    check("5 union rows on the canonical route/county/PM identities "
          "(decimal-canonical postmiles since CMP-AUD-006)",
          set(by_pm) == {"001 / ORA / 0.606", "001 / ORA / 1",
                         "001 / ORA / 1.5", "001 / ORA / 2",
                         "002 / ORA / 10"})

    # The COMPARED HG diff at PM 1.000 carries the ≠ marker.
    hg_col = header.index("HG")
    pm1 = by_pm["001 / ORA / 1"]
    check("compared HG difference shows the diff marker", DIFF in pm1[hg_col])

    # The CONTEXT 'Ramp Type' column NEVER carries a diff marker, and SHOWS the TSN value.
    rt_col = header.index("Ramp Type")
    check("context 'Ramp Type' never shows a diff marker in any row",
          all(DIFF not in r[rt_col] for r in rows))
    check("context 'Ramp Type' coalesces to the TSN value (M)",
          by_pm["001 / ORA / 2"][rt_col] == "M")
    dist_col = header.index("District")
    check("District compared and equal on the matched rows (no diff marker)",
          all(DIFF not in r[dist_col] for r in rows))

    # Total diff cells: count the ≠ across the whole Comparison body, and confirm
    # NONE of them are in the four context columns.
    ctx_cols = [header.index(c) for c in ("Ramp Name", "On/Off", "Ramp Type", "ADT")]
    total = sum(1 for r in rows for c in r if DIFF in c)
    ctx_diffs = sum(1 for r in rows for i in ctx_cols if DIFF in r[i])
    check("zero diff cells fall in the context columns", ctx_diffs == 0)
    check("at least the one compared HG diff is counted", total >= 1)
    print(f"      (union rows={len(rows)}, total diff cells={total}, context diff cells={ctx_diffs})")


def test_pm_identity_canon():
    """CMP-AUD-006: the notes' contract says '9.6' and '009.600' identify the
    SAME ramp, but the norm_pm identity split them into one-sided rows. The
    physical identity now hashes the DECIMAL-canonical postmile
    (compare_tsn_common.decimal_pm) while each side's norm_pm text stays the
    display payload."""
    print("CMP-AUD-006 — trailing-zero PM variants are ONE physical identity:")
    import compare_tsn_common as ctc
    check("decimal_pm unifies 9.6 / 9.600 / 009.600",
          ctc.decimal_pm("9.6") == ctc.decimal_pm("9.600")
          == ctc.decimal_pm("009.600") == "9.6")
    check("decimal_pm unifies the zero variants to '0' (never blank)",
          ctc.decimal_pm("0") == ctc.decimal_pm("0.0")
          == ctc.decimal_pm("000.000") == "0")
    check("decimal_pm keeps real fractions ('005.870' -> '5.87')",
          ctc.decimal_pm("005.870") == "5.87")
    k1 = rd._physical_pm_key("101", "DN", "9.6",
                             (("postmile", "9.6"),), "probe A")
    k2 = rd._physical_pm_key("101", "DN", "009.600",
                             (("postmile", "009.600"),), "probe B")
    check("the two printed variants build EQUAL physical identities",
          k1.physical_identity == k2.physical_identity)
    check("...while each keeps its own normalized display payload",
          str(k1) == "9.6" and str(k2) == "9.600")
    k3 = rd._physical_pm_key("101", "DN", "9.61",
                             (("postmile", "9.61"),), "probe C")
    check("a genuinely different postmile still differs",
          k3.physical_identity != k1.physical_identity)


def test_two_county_and_v3_refusal():
    print("county-aware identity + stale-library refusal:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_rd_d4_"))
    tsmis_path = root / "t.xlsx"
    tsn_path = root / "n.xlsx"
    out_path = root / "c.xlsx"
    # The SAME route+PM in two counties with the descriptions swapped between
    # physical locations: the D4 key must yield TWO paired rows with exactly two
    # Description differences — never a "match" that pairs across counties.
    _write_tsmis(tsmis_path, [
        _tsmis_row("101", "01-DN-101", "R", "001.000", "01/01/2026", "D", "Y", "A", "U", "101/ALPHA"),
        _tsmis_row("101", "07-LA-101", "R", "001.000", "01/01/2026", "D", "Y", "A", "U", "101/BETA"),
    ])
    _write_tsn(tsn_path, [
        ["101", "R", "1.000", "01", "2026-01-01", "D", "Y", "A", "U", "BETA", "", "", "", "", "01", "DN", ""],
        ["101", "R", "1.000", "07", "2026-01-01", "D", "Y", "A", "U", "ALPHA", "", "", "", "", "07", "LA", ""],
    ])
    res = rd.compare(tsmis_path, tsn_path, out_path, events=Events(),
                     confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows = _comparison(out_path)
    pm_col, desc_col = header.index("PM"), header.index("Description")
    by_pm = {r[pm_col]: r for r in rows}
    check("two county-distinct identities, both paired",
          set(by_pm) == {"101 / DN / 1", "101 / LA / 1"})
    check("the physical swap surfaces as TWO Description differences",
          all(DIFF in by_pm[k][desc_col] for k in by_pm))
    check("...and exactly two differing cells total",
          res.comparison_outcome.counts.differing_cells == 2)

    # A pre-v4 normalized library (no District / PM-Suffix columns) refuses.
    stale = root / "stale.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = rd.NORMALIZED_SHEET
    old_header = ["Route"] + [h for h in rd.SHARED_HEADER if h != "District"] + [
        "TSN District", "TSN County"]
    ws.append(old_header)
    ws.append(["001", "R", "0.606", "1976-02-25", "D", "Y", "A", "U", "X",
               "", "", "", "", "12", "ORA"])
    wb.save(stale)
    wb.close()
    try:
        rd._load_tsn(stale)
        check("a v3 library refuses with a rebuild hint", False)
    except ValueError as e:
        # CMP-AUD-033: the exact-header-prefix gate now catches the missing
        # District column (a pre-county-aware shape) before the marker gate.
        check("a v3 (District-less) library refuses with a rebuild hint",
              "column layout does not match" in str(e) and "rebuild" in str(e))

    # CMP-AUD-037: a CURRENT-shape library (District + the sidecars) that carries
    # no normalization marker, or a stale one, is refused on the direct path —
    # the shape gate can't see it, so the marker version is the authoritative
    # freshness check.
    nomark = root / "nomarker.xlsx"
    _write_tsn(nomark, [], marker_version=None)
    try:
        rd._load_tsn(nomark)
        check("a marker-less current-shape library refuses (CMP-AUD-037)", False)
    except ValueError as e:
        check("a marker-less current-shape library refuses (CMP-AUD-037)",
              "older TSN converter" in str(e) and "rebuild" in str(e))
    old = root / "stale-marker.xlsx"
    _write_tsn(old, [], marker_version=rd.NORMALIZATION_VERSION - 1)
    try:
        rd._load_tsn(old)
        check("a stale-marker library refuses (CMP-AUD-037)", False)
    except ValueError as e:
        check("a stale-marker library refuses (CMP-AUD-037)",
              "older TSN converter" in str(e))
    # ...and the current marker is accepted (the round-trip closes green).
    good = root / "good.xlsx"
    _write_tsn(good, [])
    try:
        rows, has_route = rd._load_tsn(good)
        check("a current-marker library is accepted", rows == [] and has_route is True)
    except ValueError:
        check("a current-marker library is accepted", False)


def main():
    test_schema()
    test_end_to_end()
    test_pm_identity_canon()
    test_two_county_and_v3_refusal()
    test_dual_layout_loader()
    test_consolidator_completion_truth()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-RAMP-DETAIL-TSN CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
