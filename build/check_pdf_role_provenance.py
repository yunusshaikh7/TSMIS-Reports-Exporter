"""CMP-AUD-066 (the PDF-role halves): a "TSMIS (PDF)" comparison side must
actually BE a PDF-sourced workbook, and a "TSMIS (Excel)" side must not be.

The PDF-vs-Excel self-check flavors used to validate only a workbook SHAPE the
Excel rendering also has, so two copies of an Excel-consolidated workbook were
accepted as `TSMIS (PDF)` vs `TSMIS (Excel)` and certified a match (Highway
Log / Highway Sequence / Highway Detail / Intersection Detail; Ramp Detail's
richer 13-column PDF header already rejects the Excel shape structurally).

Now every workbook this app writes FROM PDFs — the five per-route converted
files AND the combined conversion workbooks — carries a very-hidden versioned
`TSMIS PDF Conversion` marker sheet (pdf_table_lib.write_pdf_source_marker /
the `pdf_source_marker=` opt-in on write_route_workbook; the TSN Highway Log
consolidator shares write_route_workbook and stays UNMARKED — it is not a
TSMIS PDF conversion). The four vulnerable families' flavors enforce roles at
load time:

  * the `TSMIS (PDF)` side REQUIRES the marker (an unmarked pick refuses with
    a re-consolidate hint — pre-marker workbooks re-consolidate once);
  * the `TSMIS (Excel)` side REJECTS the marker (a PDF-sourced workbook can't
    stand in for the Excel export);
  * TSN sides keep their own normalization-marker gates (v4/v5) unchanged.

CI-safe: pure Python fixtures, no local data.
"""
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import Workbook, load_workbook  # noqa: E402

import compare_highway_detail_pdf as hd_pdf  # noqa: E402
import compare_highway_log_pdf as hl_pdf  # noqa: E402
import compare_highway_sequence_pdf as hsl_pdf  # noqa: E402
import compare_intersection_detail_pdf as id_pdf  # noqa: E402
import highway_detail_columns as hdc  # noqa: E402
import highway_log_columns as hlc  # noqa: E402
import intersection_detail_columns as idc  # noqa: E402
import pdf_table_lib  # noqa: E402
from events import Events  # noqa: E402

failures = []


def check(label, cond, detail=""):
    print(("OK   " if cond else "FAIL ") + label
          + (f" — {detail}" if detail and not cond else ""))
    if not cond:
        failures.append(label)


MARKER_SHEET = "TSMIS PDF Conversion"


def mark(path):
    """Stamp the marker the production writer must emit (shape-pinned below)."""
    wb = load_workbook(path)
    ws = wb.create_sheet(MARKER_SHEET)
    ws["A1"] = MARKER_SHEET
    ws["A2"] = 1
    ws.sheet_state = "veryHidden"
    wb.save(path)
    wb.close()


def write_wb(path, sheet, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    wb.close()


# --------------------------------------------------------------------------- #
# fixtures: minimal valid consolidated workbooks per family
# --------------------------------------------------------------------------- #

def hl_fixture(path):
    """A per-route Highway Log workbook (31 columns, route token in the name)."""
    write_wb(path, "Highway Log", hlc.HEADER,
             [["000.100"] + [""] * 30, ["000.200"] + [""] * 30])


def hsl_fixture(path):
    """A CONSOLIDATED Highway Sequence workbook (leading Route column)."""
    header = ["Route", "County", "City", "PR", "PM", "SFX", "HG", "FT",
              "Distance", "Description"]
    write_wb(path, "Highway Locations", header,
             [["001", "ALP", "", "", "001.000", "", "D", "H", "", "A ROW"]])


def hd_fixture(path):
    write_wb(path, "Highway Detail",
             ["Route"] + list(hdc.HEADER),
             [["001"] + ["000.100"] + [""] * 33])


def id_fixture(path):
    write_wb(path, "Intersection Detail",
             ["Route"] + list(idc.HEADER),
             [["001", "", "000.100", "", "04 SOL 001"] + [""] * 31])


FLAVORS = [
    ("highway_log", hl_pdf.TSMIS_PDF_VS_EXCEL, hl_fixture,
     "highway_log route 7 {}.xlsx"),
    ("highway_sequence", hsl_pdf.TSMIS_PDF_VS_EXCEL, hsl_fixture,
     "highway_sequence {}.xlsx"),
    ("highway_detail", hd_pdf.TSMIS_PDF_VS_EXCEL, hd_fixture,
     "highway_detail {}.xlsx"),
    ("intersection_detail", id_pdf.TSMIS_PDF_VS_EXCEL, id_fixture,
     "intersection_detail {}.xlsx"),
]


def run(flavor, a, b, out):
    return flavor.compare(str(a), str(b), str(out), events=Events(),
                          confirm_overwrite=lambda _p: True)


def role_pins():
    for key, flavor, fixture, name_pat in FLAVORS:
        print(f"{key} — TSMIS PDF vs Excel roles:")
        tmp = Path(tempfile.mkdtemp(prefix=f"tsmis_role_{key}_"))
        try:
            a = tmp / name_pat.format("pdf")
            b = tmp / name_pat.format("excel")
            out = tmp / "cmp.xlsx"
            fixture(a)
            fixture(b)

            # R1: an UNMARKED workbook on the PDF side refuses.
            res = run(flavor, a, b, out)
            check(f"{key}: an unmarked PDF-side pick refuses with the marker hint",
                  res.status == "error" and "PDF" in res.message
                  and ("consolidate" in res.message.lower()
                       or "conversion" in res.message.lower()),
                  f"{res.status}: {res.message[:200]}")

            # R2: a MARKED (PDF-sourced) workbook on the EXCEL side refuses.
            mark(a)
            mark(b)
            res = run(flavor, a, b, out)
            check(f"{key}: a PDF-sourced workbook on the Excel side refuses",
                  res.status == "error" and "Excel" in res.message,
                  f"{res.status}: {res.message[:200]}")

            # R3: the honest pair (marked PDF side, unmarked Excel side) runs.
            b2 = tmp / name_pat.format("excel2")
            fixture(b2)
            res = run(flavor, a, b2, out)
            check(f"{key}: the honest pair still compares",
                  res.status == "ok", f"{res.status}: {res.message[:200]}")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


def pdf_vs_tsn_pins():
    """The vs-TSN flavors' TSMIS (PDF) side requires the marker too (their TSN
    sides keep their own v4/v5 gates — pinned elsewhere). Loader-level: the
    refusal must arrive before any TSN-side work."""
    print("TSMIS (PDF) vs TSN — the PDF side requires the marker:")
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_role_tsn_"))
    try:
        a = tmp / "highway_log route 7 pdf.xlsx"
        b = tmp / "highway_log route 7 tsn.xlsx"
        hl_fixture(a)
        hl_fixture(b)
        res = run(hl_pdf.TSMIS_PDF_VS_TSN, a, b, tmp / "cmp.xlsx")
        check("HL PDF-vs-TSN: an unmarked PDF side refuses (before TSN gating)",
              res.status == "error" and "PDF" in res.message,
              f"{res.status}: {res.message[:200]}")

        a2 = tmp / "highway_detail pdf.xlsx"
        b2 = tmp / "highway_detail tsn.xlsx"
        hd_fixture(a2)
        hd_fixture(b2)
        res = run(hd_pdf.TSMIS_PDF_VS_TSN, a2, b2, tmp / "cmp2.xlsx")
        check("HD PDF-vs-TSN: an unmarked PDF side refuses",
              res.status == "error" and "PDF" in res.message,
              f"{res.status}: {res.message[:200]}")

        a3 = tmp / "intersection_detail pdf.xlsx"
        b3 = tmp / "intersection_detail tsn.xlsx"
        id_fixture(a3)
        id_fixture(b3)
        res = run(id_pdf.TSMIS_PDF_VS_TSN, a3, b3, tmp / "cmp3.xlsx")
        check("ID PDF-vs-TSN: an unmarked PDF side refuses",
              res.status == "error" and "PDF" in res.message,
              f"{res.status}: {res.message[:200]}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def writer_pins():
    """The production marker writer + the opt-in seam + the fail-closed reader."""
    print("the marker writer:")
    fn = getattr(pdf_table_lib, "write_pdf_source_marker", None)
    state = getattr(pdf_table_lib, "pdf_source_marker_state", None)
    if fn is None or state is None:
        check("pdf_table_lib.write_pdf_source_marker / pdf_source_marker_state exist",
              False)
        return
    tmp = Path(tempfile.mkdtemp(prefix="tsmis_role_writer_"))
    try:
        # write_route_workbook stamps the marker ONLY when asked (the TSN
        # Highway Log consolidator shares this writer and must stay unmarked).
        p = tmp / "route.xlsx"
        pdf_table_lib.write_route_workbook(
            [["000.100", "x"]], p, sheet_name="S", header=["PM", "V"],
            pdf_source_marker=True)
        check("write_route_workbook(pdf_source_marker=True) stamps v1",
              state(p) == 1)
        wb = load_workbook(p)
        check("...the marker sheet is very-hidden",
              MARKER_SHEET in wb.sheetnames
              and wb[MARKER_SHEET].sheet_state == "veryHidden")
        wb.close()

        p2 = tmp / "route_plain.xlsx"
        pdf_table_lib.write_route_workbook(
            [["000.100", "x"]], p2, sheet_name="S", header=["PM", "V"])
        check("the default stays UNMARKED (the TSN consolidator's path)",
              state(p2) == 0)

        # The check's own mark() must equal the production shape, so every
        # fixture in this file models the real artifact.
        p3 = tmp / "manual.xlsx"
        write_wb(p3, "S", ["PM"], [["000.100"]])
        mark(p3)
        check("the fixture marker parses at the same version", state(p3) == 1)

        # A present-but-malformed marker fails CLOSED on both roles: -1 is
        # never valid for the PDF side and never clean for the Excel side.
        p4 = tmp / "malformed.xlsx"
        write_wb(p4, "S", ["PM"], [["000.100"]])
        wb = load_workbook(p4)
        ws = wb.create_sheet(MARKER_SHEET)
        ws["A1"] = "something else"
        wb.save(p4)
        wb.close()
        check("a malformed marker reads -1 (fails closed both ways)",
              state(p4) == -1)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


writer_pins()
role_pins()
pdf_vs_tsn_pins()

if failures:
    print(f"\nFAILED: {len(failures)}")
    sys.exit(1)
print("\nPDF role provenance (CMP-AUD-066, PDF-role halves): PASS")
