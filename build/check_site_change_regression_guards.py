"""Executable guards for the two NO-FIX site/source findings (PCOA-FINAL-021/022).

Both were validated as CORRECT behaviour and recorded as prose "must not
regress" notes. Prose is not a gate, so this turns each into a check that fails
against a deliberate regression.

**022 — the site changed its prints and the parsers absorbed it.** Two changes
appear in the frozen archive and must BOTH keep working:

  1. a stray leading `GENERATE` line is now the first text line of the
     `ramp_summary`, `ramp_detail_pdf` and `intersection_detail_pdf` prints
     (the finding names a fourth, `intersection_summary_pdf` — it is EXPORT-ONLY
     and has no parser at all, which PCOA-FINAL-018 now declares in the catalog;
     asserted below rather than assumed, so a parser added later is guarded);
  2. the Highway Sequence Listing (PDF) print was re-skinned from
     `California Department of Transportation / Highway Sequence Listing` to the
     TASAS layout, with a WIDER text measure. The parser survives it because it
     derives each page's column windows from that page's own header-word
     positions — so the guard builds the SAME logical rows at both measures and
     requires identical output, which is the property that must not regress.

**021 — two genuine PDF-only rows.** Route `074` @ `000.000` occurrence 2 (prior
7.9 raw PDF page 7 line 31) and route `101` @ `R022.828` (page 142 line 23) exist
in the raw Highway Log PDF and NOT in its Excel sibling. The PDF-derived universe
must keep them and the Excel-derived universe must never synthesize them: they
must surface as ONE-SIDED, never paired against an invented partner. Witness:
`hotfix-bundles/HF-11/witness/pdf_only_rows.json`.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_site_change_regression_guards.py
"""
import json
import logging
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
sys.path.insert(0, os.path.dirname(__file__))

import compare_env
import consolidate_highway_log as hl
import consolidate_tsmis_highway_sequence_pdf as hslpdf
import highway_log_columns as hlc
import report_catalog
from _hl_fixture_pdf import make_pdf
from events import Events
from openpyxl import Workbook, load_workbook

logging.disable(logging.CRITICAL)

_fail = []

WITNESS = (Path(__file__).resolve().parent.parent
           / "docs" / "planning" / "post-comparison-perfection-output-audit"
           / "hotfix-bundles" / "HF-11" / "witness" / "pdf_only_rows.json")

# 022: the four print families the stray leading `GENERATE` line precedes.
GENERATE_FAMILIES = ("ramp_summary", "ramp_detail_pdf", "intersection_detail_pdf",
                     "intersection_summary_pdf")


def check(name, cond, detail=""):
    suffix = f"  -> {detail}" if (not cond and detail) else ""
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}{suffix}")
    if not cond:
        _fail.append(name)


# --------------------------------------------------------------------------- #
# 022a — the Highway Sequence print parses the same at BOTH text measures
# --------------------------------------------------------------------------- #
# One logical page of Highway Sequence data, laid out from the SAME window
# formulas the parser derives from each page's own header positions
# (`_boundaries`). Every data token is centred in its window, so the identical
# document can be rendered at any text measure and must parse identically —
# which is exactly the property the TASAS re-skin exercised.
_HEADER_FRACS = {                    # column -> (header label, x fraction)
    "county": ("COUNTY", 0.02),
    "city": ("CITY", 0.13),
    "pm": ("PM", 0.30),
    "hg": ("HG", 0.44),
    "ft": ("FT", 0.51),
    "dist": ("NEXT", 0.58),
    "desc": ("DESCRIPTION", 0.70),
}
_HSL_ROWS = [
    # county, city, pm, hg, ft, next, description
    ("DN", "", "001.000", "D", "H", "0.500", "JCT RTE 199"),
    ("DN", "CC", "001.500", "U", "I", "1.250", "FRONT ST"),
    ("SIS", "", "002.750", "R", "R", "0.000", "COUNTY LINE"),
]
# Helvetica 8pt averages ~0.556 em per glyph; only used to centre a token in its
# window, so the estimate only has to be close.
_GLYPH_W = 4.45


def _text_width(text):
    return _GLYPH_W * len(text)


def _hsl_page(measure, skin_title, left=40.0):
    """One data page at the given text `measure` (points), under `skin_title`'s
    banner lines."""
    def hx(frac):
        return left + frac * measure

    head = {key: (label, hx(frac), hx(frac) + _text_width(label))
            for key, (label, frac) in _HEADER_FRACS.items()}
    city_x0, pm_x0, pm_x1 = head["city"][1], head["pm"][1], head["pm"][2]
    hg_x0, ft_x0, ft_x1 = head["hg"][1], head["ft"][1], head["ft"][2]
    desc_x0 = head["desc"][1]
    # The parser's own boundaries, in the parser's own order.
    b = {"county_city": city_x0 - 4, "city_prefix": city_x0 + 30,
         "prefix_pm": pm_x0 - 10, "pm_suffix": pm_x1 + 12,
         "suffix_hg": hg_x0 - 4, "hg_ft": ft_x0 - 4,
         "ft_dist": ft_x1 + 8, "dist_desc": desc_x0 - 6}
    centres = {
        "county": (left + b["county_city"]) / 2,
        "city": (b["county_city"] + b["city_prefix"]) / 2,
        "pm": (b["prefix_pm"] + b["pm_suffix"]) / 2,
        "hg": (b["suffix_hg"] + b["hg_ft"]) / 2,
        "ft": (b["hg_ft"] + b["ft_dist"]) / 2,
        "dist": (b["ft_dist"] + b["dist_desc"]) / 2,
    }
    # The Description is LEFT-ALIGNED at the start of its window, like the real
    # print — centring a long description would push its first word back over the
    # boundary into the Next-Point column.
    desc_x0 = b["dist_desc"] + 8

    runs = []
    for i, text in enumerate(skin_title):
        runs.append((hx(0.25), 20 + i * 12, text))
    # The banner line the parser reads the document's own route claim from.
    runs.append((left, 56, "District: 01 Route: 001 Direction: S-N"))
    for _key, (label, x0, _x1) in head.items():
        runs.append((x0, 72, label))
    for i, row in enumerate(_HSL_ROWS):
        top = 92 + i * 12
        for value, key in zip(row, _HEADER_FRACS):
            if not value:
                continue
            x0 = (desc_x0 if key == "desc"
                  else centres[key] - _text_width(value) / 2)
            runs.append((x0, top, value))
    return runs


_OLD_SKIN = ("California Department of Transportation", "Highway Sequence Listing")
_TASAS_SKIN = ("TASAS", "Traffic Accident Surveillance and Analysis System",
               "HIGHWAY SEQUENCE LISTING (W/CITIES)")


def _parse_hsl(path):
    rows, stats = hslpdf.parse_pdf(str(path), Events())
    return rows, stats


def test_highway_sequence_both_skins(tmp):
    print("022a: the Highway Sequence print parses identically at BOTH measures:")
    old_pdf, new_pdf = tmp / "hsl_old.pdf", tmp / "hsl_tasas.pdf"
    make_pdf(old_pdf, [_hsl_page(360.0, _OLD_SKIN)])
    make_pdf(new_pdf, [_hsl_page(520.0, _TASAS_SKIN)])   # the wider TASAS measure
    old_rows, old_stats = _parse_hsl(old_pdf)
    new_rows, new_stats = _parse_hsl(new_pdf)
    check("the pre-re-skin layout yields every data row",
          len(old_rows) == len(_HSL_ROWS), f"{len(old_rows)} rows: {old_rows}")
    check("the TASAS layout yields every data row",
          len(new_rows) == len(_HSL_ROWS), f"{len(new_rows)} rows: {new_rows}")
    check("both layouts yield the SAME rows", old_rows == new_rows,
          f"old={old_rows}\n         new={new_rows}")
    check("neither layout leaves an unclassified line or a stray fragment",
          not old_stats["unclassified"] and not old_stats["stray_frags"]
          and not new_stats["unclassified"] and not new_stats["stray_frags"],
          f"old={old_stats} new={new_stats}")
    check("both layouts read the document's own route claim",
          list(old_stats["doc_routes"]) == list(new_stats["doc_routes"]) == ["001"],
          f"old={old_stats['doc_routes']} new={new_stats['doc_routes']}")


# --------------------------------------------------------------------------- #
# 022b — a leading GENERATE line is never data
# --------------------------------------------------------------------------- #
def test_leading_generate_line_is_ignored(tmp):
    print("022b: a stray leading GENERATE line is never read as data:")
    plain = tmp / "hsl_plain.pdf"
    generated = tmp / "hsl_generate.pdf"
    page = _hsl_page(520.0, _TASAS_SKIN)
    make_pdf(plain, [page])
    make_pdf(generated, [[(20.0, 8, "GENERATE")] + page])
    without, _ = _parse_hsl(plain)
    with_line, stats = _parse_hsl(generated)
    check("the GENERATE line changes nothing", without == with_line,
          f"without={without}\n         with={with_line}")
    check("...and is not counted as an unclassified line",
          not stats["unclassified"] and not stats["stray_frags"], repr(stats))
    check("no parsed cell anywhere contains the word GENERATE",
          not any("GENERATE" in str(v) for row in with_line for v in row),
          repr(with_line))

    print("...and every family the finding names either has a parser or none to guard:")
    consolidators = report_catalog.consolidator_by_export_subdir()
    export_only = set(report_catalog.export_only_keys())
    for family in GENERATE_FAMILIES:
        has_parser = hasattr(consolidators.get(family), "parse_pdf")
        check(f"{family}: has a PDF parser, or is declared export-only",
              has_parser or family in export_only,
              "it has no parser and is not declared export-only — a print family "
              "with neither is unguarded and unexplained")


# --------------------------------------------------------------------------- #
# 021 — the two PDF-only Highway Log rows stay one-sided
# --------------------------------------------------------------------------- #
def _hl_row(location, description):
    values = [""] * len(hlc.HEADER)
    values[0] = location
    values[hlc.HEADER.index("Description")] = description
    return values


def _write_hl_route(folder, route, rows):
    folder.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = hl.SHEET_NAME
    ws.append(list(hlc.HEADER))
    for r in rows:
        ws.append(r)
    wb.save(folder / f"highway_log_route_{route}.xlsx")


def _status_counts(out_path):
    """Every row's Status from the produced comparison workbook."""
    wb = load_workbook(out_path, read_only=True, data_only=True)
    try:
        ws = wb["Comparison"]
        header = [str(c) if c is not None else "" for c in next(
            ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = header.index("Status")
        counts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            value = row[idx] if idx < len(row) else None
            if value:
                counts[str(value)] = counts.get(str(value), 0) + 1
        return counts
    finally:
        wb.close()


def test_pdf_only_rows_stay_one_sided(tmp):
    print("021: the two PDF-only Highway Log rows stay one-sided, never synthesized:")
    witness = json.loads(WITNESS.read_text(encoding="utf-8"))
    check("the committed witness names both rows", len(witness["rows"]) == 2,
          repr(witness))

    base = tmp / "hl021"
    pdf_side, excel_side = base / "PDF", base / "EXCEL"
    for row in witness["rows"]:
        route = row["route"]
        shared = [_hl_row(row["location"], "SHARED ROW")]
        # The PDF side carries the extra occurrence; the Excel side does not.
        pdf_rows = shared + [_hl_row(row["location"], row["description"])] \
            if row["pdf_count"] > row["excel_count"] else shared
        _write_hl_route(pdf_side / "highway_log", route, pdf_rows)
        _write_hl_route(excel_side / "highway_log", route, shared)

    out = base / "cmp.xlsx"
    res = compare_env.HIGHWAY_LOG.compare_folders(
        str(pdf_side), str(excel_side), str(out), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    check("the comparison ran", res.status == "ok",
          f"{res.status!r} {res.message!r}")
    if res.status != "ok":
        return
    counts = _status_counts(out)
    # compare_core labels a one-sided row "<side label> only"; the side labels
    # come from the two folder names, PDF and EXCEL.
    pdf_only = sum(v for k, v in counts.items() if k.strip().lower() == "pdf only")
    excel_only = sum(v for k, v in counts.items() if k.strip().lower() == "excel only")
    check("each PDF-only row is reported as ONE-SIDED, not paired",
          pdf_only == len(witness["rows"]),
          f"{pdf_only} PDF-only of {len(witness['rows'])} expected; statuses={counts}")
    check("no PDF-only row was synthesized into the Excel side",
          excel_only == 0, f"statuses={counts}")


def main():
    print("=== site-change and source-universe regression guards "
          "(PCOA-FINAL-021/022) ===")
    with tempfile.TemporaryDirectory(prefix="tsmis_guards_") as tmp:
        root = Path(tmp)
        test_highway_sequence_both_skins(root)
        test_leading_generate_line_is_ignored(root)
        test_pdf_only_rows_stay_one_sided(root)
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL SITE-CHANGE REGRESSION GUARDS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
