"""Golden check for the TSMIS-vs-TSN Highway Detail comparator
(scripts/compare_highway_detail_tsn.py) — the FLAT recipe (route + canonical PM).

Locks: the CompareSchema wiring (canonical Post Mile key at column 0; NO context
fields — every shared column compared and counted; the Med V/WDA rule; the Notes
legend_writer), the canonical roadbed-aware Post Mile key (TSMIS trailing R/L ==
TSN bare PM + HG, the equation marker excluded from the key and compared as the
PS column), the NA 'A'->blank / zero-padding / length / Med-WDA / route-token
normalizations, the position-based TSMIS-consolidated loader, and end-to-end that
a normalization still produces a MATCH while everything present in both systems
IS counted — the RU-Eff-vs-BEG_DATE structural column and a genuine lane change
both flag. No Excel; CI-safe (synthetic fixtures modeled on the real statewide
bundle the family was reconciled against).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_highway_detail_tsn.py
"""
import inspect
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_highway_detail_tsn as hdt
import compare_tsn_common as ctc
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []
DIFF = " ≠ "


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _write_tsmis(path, rows):
    """Synthetic CONSOLIDATED Highway Detail: 'Route' + the 34 export columns.
    The loader reads by POSITION, but CMP-AUD-034 now binds the EXACT header, so
    the fixture uses the real one (row VALUES are still positional — no assertion
    changes)."""
    wb = Workbook()
    ws = wb.active
    ws.title = hdt.TSMIS_SHEET
    ws.append(list(hdt._TSMIS_HEADER))
    for r in rows:
        ws.append(r + [None] * (35 - len(r)))
    wb.save(path)
    wb.close()


def _tsmis_row(route, pm, length, dor, hg, ac, aceff, city, ru, rueff, desc, na,
               lbeff=None, lbt=None, lbln=None, lbf=None, lbto1=None, lbtr1=None,
               lbwid=None, lbto2=None, lbtr2=None,
               medeff=None, medt=None, medc=None, medb=None, wda=None,
               rbeff=None, rbt=None, rbln=None, rbf=None, rbto1=None, rbtr1=None,
               rbwid=None, rbto2=None, rbtr2=None):
    return [route, pm, length, dor, hg, ac, aceff, city, ru, rueff, desc, na,
            lbeff, lbt, lbln, lbf, lbto1, lbtr1, lbwid, lbto2, lbtr2,
            medeff, medt, medc, medb, wda,
            rbeff, rbt, rbln, rbf, rbto1, rbtr1, rbwid, rbto2, rbtr2]


_TSN_COLS = ["THY_ID", "DIST", "CNTY", "RTE", "RTE_SFX", "DIST_CNTY_ROUTE", "PP",
             "POSTMILE", "E_IND", "LENGTH", "REC_DATE", "HG", "AC", "ACC_SIG",
             "ACC_EFF_DATE", "CITY", "POP_CODE", "BEG_DATE", "ADT_AMT", "PROFILE",
             "BREAK_DESC", "LK_BACK_ADT", "CHNGMILE", "DVM", "DESCRIPTION",
             "NON_ADD", "LT_SIG", "L_EFF_DATE", "L_ST", "L_NO_LANES", "L_SF",
             "L_OT_TOT", "L_OT_TR", "L_TR_WID", "L_IN_TOT", "L_IN_TR", "MED_SIG",
             "M_EFF_DATE", "M_TYPE_CODE", "M_CL", "M_BA", "M_WID", "M_VA",
             "RT_SIG", "R_EFF_DATE", "R_ST", "R_NO_LANES", "R_SF", "R_IN_TOT",
             "R_IN_TR", "R_TR_WID", "R_OT_TOT", "R_OT_TR", "SEG_ORDER_ID",
             "REFERENCE_DATE", "EXTRACT_DATE"]


def _write_tsn(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = hdt.TSN_SHEET
    ws.append(_TSN_COLS)
    for r in rows:
        ws.append([r.get(c) for c in _TSN_COLS])
    wb.save(path)
    wb.close()


def _comparison(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Comparison"]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
        return header, rows, wb.sheetnames
    finally:
        wb.close()


def test_schema():
    print("schema + normalizers:")
    sc = hdt._SCHEMA
    check("key is the canonical Post Mile (column 0)",
          sc.header[sc.key_field] == "Post Mile" and sc.key_field == 0)
    check("side names TSMIS / TSN", sc.side_a == "TSMIS" and sc.side_b == "TSN")
    check("position-aligned: NO context fields (nothing suppressed or greyed)",
          tuple(sc.context_fields) == ())
    check("Med V/WDA uses the Med-Wid zero-pad rule", sc.medwid_fields == ("Med V/WDA",))
    check("the 6 date columns + RU Eff ARE compared (nothing suppressed)",
          all(f in sc.header and f not in sc.context_fields for f in hdt.DATE_FIELDS))
    check("Notes legend_writer set (documents the normalizations)",
          sc.legend_writer is not None)
    check("base schema is clean; the Report View closure is added per-call",
          sc.extra_sheet_writer is None and not sc.report_view_diff_check)
    check("35 shared columns (Post Mile + PS + the 33 remaining export columns)",
          len(hdt.SHARED_HEADER) == 35)
    # The canonical roadbed-aware key.
    check("TSMIS glued roadbed suffix is the key ('000.080R')",
          hdt.pm_canon("000.080R", "R") == "000.080R")
    check("TSN bare PM + HG=R keys identically",
          hdt.pm_canon("000.080", "R") == "000.080R")
    check("HG supplies the roadbed when the TSMIS marker slot holds 'E' "
          "('000.137E' + HG=L)", hdt.pm_canon("000.137E", "L") == "000.137L")
    check("the equation marker is NOT part of the key ('C043.925' + E + HG=R "
          "== TSMIS 'C043.925R')",
          hdt.pm_canon("C043.925", "R") == hdt.pm_canon("C043.925R", "D") == "C043.925R")
    check("plain divided-highway PM keys bare", hdt.pm_canon("011.228", "D") == "011.228")
    check("mile zero-pads to the printed form (' 11.228' -> '011.228')",
          hdt.pm_canon(" 11.228", "D") == "011.228")
    check("PS: explicit E_IND / a glued trailing E both read 'E'",
          hdt.pm_suffix("000.000", "E") == "E" and hdt.pm_suffix("000.000E") == "E"
          and hdt.pm_suffix("000.080R") == "")

    # CMP-AUD-042: _normalized_row re-projects an ALREADY-projected library row,
    # so every field must be idempotent. PS was not: it re-ran pm_suffix() over
    # the stored marker, parsed 'E' as a glued postmile token, found no trailing
    # letters and returned '' — erasing a real difference into a clean match.
    def _reproject(post_mile, ps, hg):
        vals = [""] * (len(hdt.SHARED_HEADER) + 1)
        vals[0] = "key"
        vals[1 + hdt.SHARED_HEADER.index("Post Mile")] = post_mile
        vals[1 + hdt.SHARED_HEADER.index("PS")] = ps
        vals[1 + hdt.SHARED_HEADER.index("HG")] = hg
        out = hdt._normalized_row(vals)
        return (out[1 + hdt.SHARED_HEADER.index("Post Mile")],
                out[1 + hdt.SHARED_HEADER.index("PS")])

    check("PS: a stored 'E' survives re-projection (never silently erased)",
          _reproject("044.236", "E", "")[1] == "E")
    check("PS/Post Mile: re-projection is idempotent across blank/E x roadbed",
          all(_reproject(pm, ps, hg) == (pm, ps) for pm, ps, hg in (
              ("044.236", "E", ""), ("044.236", "", ""),
              ("R044.236R", "E", "R"), ("R044.236L", "", "L"),
              ("044.236R", "E", "R"), ("012.500R", "", "R"),
              ("L012.500", "E", ""))))
    # Normalizations.
    check("NA: TSN 'A' folds to blank; 'N' stays",
          hdt._norm_na("A") == "" and hdt._norm_na("N") == "N" and hdt._norm_na(None) == "")
    check("zero-padding: '02' == '2', '00' == '0'; blank stays blank",
          hdt._norm_num("02") == "2" and hdt._norm_num("00") == "0"
          and hdt._norm_num("") == "" and hdt._norm_num(None) == "")
    check("length: raw DB precision -> the printed 3-decimal mile "
          "(0.01098 -> '000.011')",
          hdt._norm_len(0.01098) == "000.011" and hdt._norm_len("000.055") == "000.055")
    # CMP-AUD-138: quantize the EXACT decimal, never through binary64. 0.0135 has
    # no exact float form, so float() lands below the tie and yields 000.013 where
    # the source value and the D01 print both say 000.014. Statewide census: this
    # is the ONLY one of 60,083 LENGTH values the float path got wrong.
    check("length: the exact decimal tie rounds up, not through binary64 "
          "(row 32565, '1.35E-2' -> '000.014')",
          hdt._norm_len("1.35E-2") == "000.014"
          and hdt._norm_len(0.0135) == "000.014"
          and hdt._norm_len("0.0135") == "000.014")
    check("length: the OTHER accepted allowlist row does not move "
          "(row 32564, '7.4999999999999997E-3' -> '000.007')",
          hdt._norm_len("7.4999999999999997E-3") == "000.007")
    check("length: tie neighbours, sign and carry boundaries hold",
          hdt._norm_len("0.0134") == "000.013"
          and hdt._norm_len("0.0136") == "000.014"
          and hdt._norm_len("-1.35E-2") == "-00.014"
          and hdt._norm_len("0.9995") == "001.000"
          and hdt._norm_len("99.9995") == "100.000")
    check("length: blank, non-numeric and non-finite pass through unchanged",
          hdt._norm_len("") == "" and hdt._norm_len("n/a") == "n/a"
          and hdt._norm_len("NaN") == "NaN" and hdt._norm_len(None) == "")
    check("Med WDA glue: TSN 14+'Z' == TSMIS '14Z'; '8V' pads to '08V'",
          hdt._norm_wda("14Z") == "14Z" and hdt._norm_wda("8V") == "08V")
    check("dates: a datetime cell renders to the printed YY-MM-DD text",
          hdt._norm_date(datetime(1996, 12, 2)) == "96-12-02"
          and hdt._norm_date("96-12-02") == "96-12-02")
    check("route token matches the TSMIS filenames ('5'+'S' -> '005S'; 0 -> '000')",
          hdt._norm_route_token("5", "S") == "005S"
          and hdt._norm_route_token("001") == "001"
          and hdt._norm_route_token(0) == "000")
    check("description whitespace collapses (TSN fixed-width padding)",
          hdt._norm_desc("EB 55-239              ") == "EB 55-239")
    check("Report View: a date-column diff classifies 'soft' (red, not Major); "
          "an attribute diff 'hard'",
          hdt._rv_classify("RU Eff") == "soft" and hdt._rv_classify("PS") == "soft"
          and hdt._rv_classify("LB #Ln") == "hard")
    cmp_fields = [spec[1] for column in hdt._RV_GRID
                  for spec in (column[2], column[5]) if spec[0] == "cmp"]
    asserting_fields = [sc.header[index] for index in sc.field_indices
                        if not sc.is_context(index)]
    check("Report View grid contains every asserting non-key field exactly once",
          set(cmp_fields) == set(asserting_fields)
          and len(cmp_fields) == len(asserting_fields) == len(set(cmp_fields)))


def test_report_view_typed_truth():
    print("Report View typed comparison truth:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_hd_rv_truth_"))
    out = root / "report-view.xlsx"
    # City, LB #Ln, and Length are hard differences; RU Eff and Date of Rec
    # are soft. Formula-leading/error-looking values exercise the familiar
    # view's literal-write boundary. AC is made
    # context-only in this synthetic schema to exercise the asserting gate.
    sc = hdt.dataclasses.replace(hdt._SCHEMA, context_fields=("AC",))

    def row(side_b=False):
        values = {name: "" for name in sc.header}
        values.update({
            "Post Mile": "000.204",
            "PS": "=1+1",
            "Description": "ASCII SPACE" if side_b else "  ASCII   SPACE  ",
            "City": "case" if side_b else "Case",
            "HG": "literal ≠ content",
            "NA": "#N/A",
            "Length": "+RIGHT" if side_b else "+LEFT",
            "Date of Rec": "@RIGHT" if side_b else "@LEFT",
            "Acc-Cont Eff": "-SAFE",
            "LB #Ln": 0 if side_b else None,
            "RU Eff": "74-01-01" if side_b else "73-01-01",
            "AC": "F" if side_b else "E",
        })
        return ["007"] + [values[name] for name in sc.header]

    wb = Workbook(write_only=True)
    hdt._write_report_view(
        wb,
        {"sc": sc, "rows_a": [row()], "rows_b": [row(True)]},
        [{}],
    )
    wb.save(out)
    wb.close()

    wb = load_workbook(out, data_only=True)
    try:
        ws = wb["Report View"]

        def grid_cell(line, field):
            slot = 2 if line == 0 else 5
            for grid_index, column in enumerate(hdt._RV_GRID):
                if column[slot] == ("cmp", field):
                    return ws.cell(
                        row=5 + line,
                        column=len(hdt._RV_AUX) + grid_index + 1,
                    )
            raise AssertionError(f"missing Report View field: {field}")

        desc = grid_cell(1, "Description")
        city = grid_cell(0, "City")
        literal = grid_cell(0, "HG")
        zero = grid_cell(1, "LB #Ln")
        soft = grid_cell(0, "RU Eff")
        context = grid_cell(0, "AC")
        eq_formula = grid_cell(0, "PS")
        eq_error = grid_cell(1, "NA")
        plus_diff = grid_cell(0, "Length")
        at_diff = grid_cell(0, "Date of Rec")
        minus_equal = grid_cell(0, "Acc-Cont Eff")
        check("Report View ASCII-space equality follows shared Excel TRIM",
              desc.value == "ASCII SPACE")
        check("Report View equality remains case-sensitive",
              city.value == "Case ≠ case")
        check("equal literal difference-marker content is displayed but not counted",
              literal.value == "literal ≠ content")
        check("blank-vs-zero is asserting and preserves the dot display",
              zero.value == "· ≠ 0")
        check("only asserting unequal cells render as differences",
              context.value == "E" and " ≠ " not in context.value)
        check("Report View guards =/+/−/@ and Excel-error source literals",
              eq_formula.value == "=1+1" and eq_formula.data_type == "s"
              and eq_error.value == "#N/A" and eq_error.data_type == "s"
              and plus_diff.value == "+LEFT ≠ +RIGHT" and plus_diff.data_type == "s"
              and at_diff.value == "@LEFT ≠ @RIGHT" and at_diff.data_type == "s"
              and minus_equal.value == "-SAFE" and minus_equal.data_type == "s")
        check("hard and soft differences retain the red palette",
              city.fill.fgColor.rgb[-6:] == hdt._RV_FILLS["hard"][0]
              and soft.fill.fgColor.rgb[-6:] == hdt._RV_FILLS["soft"][0]
              and literal.fill.fgColor.rgb[-6:] == hdt._RV_FILLS["eq"][0])
        data_rows = list(ws.iter_rows(min_row=5, values_only=True))
        check("matched record writes its full typed Diffs total on exactly two rows",
              len(data_rows) == 2
              and all(record[0] == 3 and record[1] == 5
                      for record in data_rows))
    finally:
        wb.close()

    source = inspect.getsource(hdt._write_report_view)
    cmp_block = source.split('if kind == "cmp":', 1)[1].split(
        'return ("", "blank")', 1)[0]
    check("Report View cmp branch has no local strip/equality truth",
          "compared_cell(sc, field_index[ref], ra, rb, off=1)" in cmp_block
          and "cell.asserting" in cmp_block and "cell.equal" in cmp_block
          and "aval(" not in cmp_block and ".strip(" not in cmp_block
          and "tm == tn" not in cmp_block)


def _tsn_base(pm, hg="D", **over):
    row = {"DIST": "11", "CNTY": "SD", "RTE": "007", "RTE_SFX": None,
           "DIST_CNTY_ROUTE": "11-SD-007", "PP": None, "POSTMILE": pm,
           "E_IND": None, "LENGTH": 0.055, "REC_DATE": "96-12-02", "HG": hg,
           "AC": "E", "ACC_EFF_DATE": "96-12-02", "CITY": None, "POP_CODE": "R",
           "BEG_DATE": "22-01-01", "ADT_AMT": " 038250", "PROFILE": "S",
           "DESCRIPTION": "BEGIN SPUR ROUTE 7    ", "NON_ADD": "A",
           "L_EFF_DATE": "96-12-02", "L_ST": "H", "L_NO_LANES": "2", "L_SF": "Z",
           "L_OT_TOT": "8", "L_OT_TR": "8", "L_TR_WID": "24", "L_IN_TOT": "8",
           "L_IN_TR": "8", "M_EFF_DATE": "96-12-02", "M_TYPE_CODE": "J",
           "M_CL": "7", "M_BA": "Z", "M_WID": 8, "M_VA": "V",
           "R_EFF_DATE": "96-12-02", "R_ST": "H", "R_NO_LANES": "2", "R_SF": "Z",
           "R_IN_TOT": "8", "R_IN_TR": "8", "R_TR_WID": "24", "R_OT_TOT": "8",
           "R_OT_TR": "8"}
    row.update(over)
    return row


def test_end_to_end():
    print("end-to-end (normalizations still match; every shared column counted):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_hd_tsn_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # PM S000.000: identical after the normalizations (lanes '02'/'2', WDA
    #   '08V'/8+'V', length text/float, TSN padded desc, NA 'N'/'N') EXCEPT the
    #   structural RU Eff vs BEG_DATE slot — exactly 1 counted diff.
    # PM 001.000: a genuine lane change (03 vs 2) + NA 'N' vs TSN add-mileage
    #   (folds to blank) + the RU Eff slot.
    _write_tsmis(tsmis, [
        _tsmis_row("007", "S000.000", "000.055", "96-12-02", "D", "E", "96-12-02",
                   None, "R", "96-12-02", "BEGIN SPUR ROUTE 7", "N",
                   "96-12-02", "H", "02", None, "08", "08", "24", "08", "08",
                   "96-12-02", "J", "7", "Z", "08V",
                   "96-12-02", "H", "02", "Z", "08", "08", "24", "08", "08"),
        _tsmis_row("007", "001.000", "000.055", "96-12-02", "D", "E", "96-12-02",
                   None, "R", "96-12-02", "BEGIN SPUR ROUTE 7", "N",
                   "96-12-02", "H", "03", "Z", "08", "08", "24", "08", "08",
                   "96-12-02", "J", "7", "Z", "08V",
                   "96-12-02", "H", "02", "Z", "08", "08", "24", "08", "08"),
    ])
    _write_tsn(tsn, [
        _tsn_base("000.000", PP="S", NON_ADD="N", L_SF=None),
        _tsn_base("001.000"),
    ])
    res = hdt.compare(tsmis, tsn, out, events=Events(),
                      confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows, sheets = _comparison(out)
    check("Notes sheet present (the normalization record)", "Notes" in sheets)
    check("Report View sheet appended (the printed two-line replica)",
          "Report View" in sheets)
    formula_wb = load_workbook(out, read_only=True, data_only=False)
    try:
        summary = formula_wb["Summary"]
        report_checks = [row[2].value for row in summary.iter_rows()
                         if len(row) >= 3
                         and row[1].value ==
                         "Report View Diffs agree with the Comparison"]
        report_formula = str(report_checks[0]) if len(report_checks) == 1 else ""
        check("Summary independently cross-checks the two-line Report View total",
              len(report_checks) == 1
              and "SUM('Report View'!B:B)=2*SUM(Comparison!" in report_formula
              and '"OK","CHECK"' in report_formula)
    finally:
        formula_wb.close()
    pm = header.index("Post Mile")
    by = {r[pm]: r for r in rows}
    check("both PMs matched (status Both)",
          by["S000.000"][header.index("Status")] == "Both"
          and by["001.000"][header.index("Status")] == "Both")
    r0, r1 = by["S000.000"], by["001.000"]
    for fld in ("NA", "LB #Ln", "Med V/WDA", "Length", "Description", "Date of Rec"):
        check(f"{fld} normalized equal — no diff on the clean row",
              DIFF not in r0[header.index(fld)])
    check("RU Eff vs BEG_DATE flags (the structural slot; compared by position)",
          DIFF in r0[header.index("RU Eff")])
    check("LB S/F blank == blank (both sides empty on the clean row)",
          DIFF not in r0[header.index("LB S/F")])
    check("clean row counts exactly its one structural diff (RU Eff)",
          r0[header.index("Diffs")] in ("1", "1.0"))
    check("a genuine lane change (03 vs 2) still flags",
          DIFF in r1[header.index("LB #Ln")])
    check("NA 'N'(TSMIS) vs TSN add-mileage 'A'(->blank) is a COUNTED diff",
          DIFF in r1[header.index("NA")])


def test_roadbed_and_ps():
    """The two encodings of an independent-alignment roadbed row must PAIR, and
    an equation-marker disagreement must flag in PS instead of splitting."""
    print("roadbed keying + the PS marker column:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_hd_rb_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # TSMIS glues the roadbed ('000.080R'); TSN prints the bare PM with HG=R.
    # TSMIS also shows no E where TSN carries E_IND='E' — PS flags, the row pairs.
    _write_tsmis(tsmis, [
        _tsmis_row("282", "000.080R", "000.055", "96-12-02", "R", "E", "96-12-02",
                   None, "U", "96-12-02", "D AVE", None,
                   "96-12-02", "H", "02", "Z", "08", "08", "24", "08", "08",
                   "96-12-02", "J", "7", "Z", "08V",
                   "96-12-02", "H", "02", "Z", "08", "08", "24", "08", "08"),
    ])
    _write_tsn(tsn, [
        _tsn_base("000.080", hg="R", RTE="282", DIST_CNTY_ROUTE="11-SD-282",
                  E_IND="E", POP_CODE="U", DESCRIPTION="D AVE"),
    ])
    res = hdt.compare(tsmis, tsn, out, events=Events(),
                      confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows, _ = _comparison(out)
    check("ONE matched row (the two roadbed encodings paired, not one-sided)",
          len(rows) == 1 and rows[0][header.index("Status")] == "Both")
    check("the canonical key shows the roadbed ('000.080R')",
          rows[0][header.index("Post Mile")] == "000.080R")
    check("PS flags the equation-marker disagreement (blank vs E)",
          DIFF in rows[0][header.index("PS")])


def test_normalized_library_idempotent():
    """A normalized TSN-library workbook re-projects at read time, so a STALE
    library (built before a normalization change) is repaired on read — and a
    fresh one is unchanged (idempotence)."""
    print("normalized-library path re-applies the normalizations:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_hd_norm_"))
    norm = root / "tsn_norm.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = hdt.NORMALIZED_SHEET
    ws.append(["Route"] + hdt.SHARED_HEADER + list(hdt._NORMALIZED_SIDECARS))

    def nrow(route, pm, hg="D", **fields):
        # The normalized shape carries the TSN District/County sidecars after the
        # shared header (CMP-AUD-033 binds the layout; _normalized_row slices to
        # the shared width, so the sidecar values are reference-only here).
        d = {"Post Mile": pm, "HG": hg}
        d.update(fields)
        return [route] + [d.get(f, "") for f in hdt.SHARED_HEADER] + ["07", "LA"]

    # A stale library that stored the RAW TSN forms: NA 'A', unpadded lanes,
    # an unpadded WDA — all must normalize on read.
    ws.append(nrow("001", "000.080", hg="R", **{"NA": "A", "LB #Ln": "02",
                                                "Med V/WDA": "8V"}))
    ctc.write_normalization_marker(wb, hdt.NORMALIZATION_VERSION)  # CMP-AUD-037 current marker
    wb.save(norm)
    wb.close()
    rows, _ = hdt._load_tsn(norm)
    r = rows[0]
    ix = {f: 1 + i for i, f in enumerate(hdt.SHARED_HEADER)}
    check("stale 'A' NA folds to blank on read", r[ix["NA"]] == "")
    check("stale '02' lanes normalize to '2'", r[ix["LB #Ln"]] == "2")
    check("stale '8V' WDA pads to '08V'", r[ix["Med V/WDA"]] == "08V")
    check("the roadbed key rebuilds from the stored HG ('000.080' + R)",
          r[ix["Post Mile"]] == "000.080R")

    # CMP-AUD-037: HD's direct loader had NO freshness gate — a normalized
    # workbook from any prior normalizer was trusted. A marker-less (or stale)
    # library now refuses with a rebuild hint; a current one is accepted.
    nomark = root / "tsn_nomarker.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = hdt.NORMALIZED_SHEET
    ws2.append(["Route"] + hdt.SHARED_HEADER + list(hdt._NORMALIZED_SIDECARS))
    ws2.append(nrow("001", "000.080", hg="R"))
    wb2.save(nomark)
    wb2.close()
    try:
        hdt._load_tsn(nomark)
        check("a marker-less library refuses (CMP-AUD-037)", False)
    except ValueError as e:
        check("a marker-less library refuses (CMP-AUD-037)",
              "older TSN converter" in str(e) and "rebuild" in str(e))


def main():
    test_schema()
    test_report_view_typed_truth()
    test_end_to_end()
    test_roadbed_and_ps()
    test_normalized_library_idempotent()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-HIGHWAY-DETAIL-TSN CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
