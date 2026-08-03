"""Golden check for the cross-environment RAMP DETAIL PM re-key
(scripts/compare_env.RAMP_DETAIL + compare_core).

Locks the headline finding of the 2026-06-16 ramp-comparison audit. Ramp Detail's
first column (a district-county-route Location / County) is COARSE — it repeats
for every ramp on a route — so the original "key on the first column" behavior
aligned rows POSITIONALLY within the route. One ramp inserted mid-route then
mis-paired every row after it, cascading into spurious field diffs: the real
delivered PROD-vs-TEST workbook reported 1,451 differing cells, ~99.4% of them
positional inflation; the TRUE difference was 8 cells / 4 rows + 10 one-sided
ramps. v0.11.0 set CompareSchema.key_field to the granular postmile ("PM") column
(compare_env.RAMP_DETAIL.key_col="PM"), collapsing the cascade to the truth.

check_compare_keyfield.py locks the generic key_field MECHANISM with a toy schema;
this check locks the REAL Ramp Detail adapter wiring end to end — that
RAMP_DETAIL keys on "PM", that the loader reads per-route "TSAR - Ramp Detail"
sheets, and that a mid-route insert produces ONE one-sided ramp with zero
spurious diffs through the actual compare_folders path.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_ramp_detail.py
"""
import os
import shutil
import sys
import tempfile
from dataclasses import replace
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
from compare_core import _DIFF_MARK, count_diffs, keys_for, union_keys
from events import Events
from openpyxl import Workbook, load_workbook

# A miniature Ramp Detail route: County is the COARSE first column (identical on
# every ramp of the route); PM is the granular postmile that actually identifies
# a ramp. Side B inserts ONE new ramp (PM 2.500) in the middle and changes
# nothing else.
ROUTE = "001"
HEADER = ["County", "PM", "Ramp ID", "Lighting"]
DATA_A = [
    ["LA", "1.000", "ON-A", "Yes"],
    ["LA", "2.000", "OFF-B", "No"],
    ["LA", "3.000", "ON-C", "Yes"],
    ["LA", "4.000", "OFF-D", "No"],
]
DATA_B = [
    ["LA", "1.000", "ON-A", "Yes"],
    ["LA", "2.000", "OFF-B", "No"],
    ["LA", "2.500", "ON-NEW", "Yes"],          # inserted mid-route
    ["LA", "3.000", "ON-C", "Yes"],
    ["LA", "4.000", "OFF-D", "No"],
]
# Consolidated shape the engine compares: [route, *per-route columns].
ROWS_A = [[ROUTE] + r for r in DATA_A]
ROWS_B = [[ROUTE] + r for r in DATA_B]

# The REAL 11-column Ramp Detail export header (CMP-AUD-032 pins the cross-env
# schema, so the compare_folders end-to-end must use the true site layout; the
# key_field unit test above stays on the toy header — it drives _schema directly,
# not the header recognizer). Location is coarse; PM is the granular key.
import compare_ramp_detail_tsn as _rd          # noqa: E402
RD_HEADER = list(_rd._TSMIS_HEADER[1:])


def _rd_row(pm, desc):
    row = [""] * len(RD_HEADER)
    row[0] = "12-ORA-001"                       # Location: district-county-route
    row[1] = "R"                                # unnamed col B: postmile prefix
    row[RD_HEADER.index("PM")] = pm
    row[RD_HEADER.index("Description")] = desc
    return row


RD_DATA_A = [_rd_row("1.000", "ON-A"), _rd_row("2.000", "OFF-B"),
             _rd_row("3.000", "ON-C"), _rd_row("4.000", "OFF-D")]
RD_DATA_B = [_rd_row("1.000", "ON-A"), _rd_row("2.000", "OFF-B"),
             _rd_row("2.500", "ON-NEW"),        # inserted mid-route
             _rd_row("3.000", "ON-C"), _rd_row("4.000", "OFF-D")]


def test_config_is_pm_keyed():
    """The audit-validated wiring: Ramp Detail keys on PM and reads the per-route
    'TSAR - Ramp Detail' sheets. A revert here is exactly the regression that
    re-inflated the diff count, so pin it explicitly."""
    rd = compare_env.RAMP_DETAIL
    assert rd.key_col == "PM", ("Ramp Detail must key on PM", rd.key_col)
    assert rd.sheet_name == "TSAR - Ramp Detail", rd.sheet_name
    assert rd.subdir == "ramp_detail", rd.subdir


def test_pm_key_collapses_coarse_cascade():
    """Through the REAL schema the adapter builds: coarse (first-column) keying
    cascades the single insert into spurious diffs; PM keying isolates it."""
    sc_pm = compare_env.RAMP_DETAIL._schema(HEADER, "SSOR-PROD", "SSOR-TEST")
    assert sc_pm.key_field == HEADER.index("PM"), \
        ("the adapter must resolve PM to its header index", sc_pm.key_field)
    sc_coarse = replace(sc_pm, key_field=0)

    def counts(sc, kf):
        kt = keys_for(ROWS_A, True, kf)
        kn = keys_for(ROWS_B, True, kf)
        return count_diffs(sc, ROWS_A, ROWS_B, kt, kn, union_keys(kt, kn), True)

    # Coarse County key: occ-3 and occ-4 mis-pair, the inserted ramp falls out
    # as occ-5 — 5 spurious differing cells across 2 rows + 1 one-sided.
    coarse = counts(sc_coarse, 0)
    assert coarse["both"] == 4 and coarse["n_only"] == 1, coarse
    assert coarse["diff_cells"] == 5 and coarse["diff_rows"] == 2, \
        ("coarse first-column key must cascade the mid-route insert", coarse)

    # PM key: the matched postmiles are identical, the one new postmile is
    # correctly one-sided — ZERO spurious diffs.
    pm = counts(sc_pm, sc_pm.key_field)
    assert pm["both"] == 4 and pm["t_only"] == 0 and pm["n_only"] == 1, pm
    assert pm["diff_cells"] == 0 and pm["diff_rows"] == 0, \
        ("PM key must isolate the insert to one one-sided ramp", pm)


def _write_route_file(path, sheet, header, data):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for row in data:
        ws.append(row)
    wb.save(path)
    wb.close()


def test_end_to_end_values_workbook():
    """Drive the actual compare_folders path over real per-route XLSX files and
    read the written VALUES workbook back: the mid-route insert must surface as
    exactly one one-sided ramp with zero differing cells."""
    root = Path(tempfile.mkdtemp())
    try:
        sheet = compare_env.RAMP_DETAIL.sheet_name
        a = root / "2026-06-16 ssor-prod" / "ramp_detail"
        b = root / "2026-06-16 ssor-test" / "ramp_detail"
        a.mkdir(parents=True)
        b.mkdir(parents=True)
        _write_route_file(a / f"ramp_detail_route_{ROUTE}.xlsx", sheet, RD_HEADER, RD_DATA_A)
        _write_route_file(b / f"ramp_detail_route_{ROUTE}.xlsx", sheet, RD_HEADER, RD_DATA_B)

        out = root / "cmp.xlsx"
        res = compare_env.RAMP_DETAIL.compare_folders(
            a.parent, b.parent, out, events=Events(),
            confirm_overwrite=lambda _p: True, mode="values")
        assert res.status == "ok", (res.status, res.message)
        assert res.verdict == "diff", res.verdict
        assert "DIFFERENCES FOUND" in res.summary_lines[0], res.summary_lines[0]

        wb = load_workbook(out, read_only=True, data_only=True)
        body = list(wb["Comparison"].iter_rows(values_only=True))[1:]
        wb.close()
        # has_route layout: A=Route B=PM C=# D=A Row E=B Row F=Status G=Diffs H..
        statuses = [r[5] for r in body]
        one_sided = [s for s in statuses if s and s != "Both"]
        assert len(body) == 5, ("union = 4 matched + 1 inserted", len(body))
        assert statuses.count("Both") == 4, statuses
        assert len(one_sided) == 1 and one_sided[0].endswith("only"), one_sided
        neq = sum(1 for r in body for v in r
                  if isinstance(v, str) and _DIFF_MARK in v)
        assert neq == 0, ("PM keying must leave zero differing cells", neq)
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_missing_key_column_fails_closed():
    """CMP-AUD-028: a CONFIGURED identity column is mandatory. It used to log and
    fall back to column 0, so two malformed key-less workbooks paired on their
    first column and returned a clean MATCH. Every keyed adapter now refuses a
    header that lacks its key column (case/whitespace-tolerant when present); an
    unkeyed adapter still uses column 0; and the Ramp Detail end-to-end returns a
    fail-closed error instead of a false match."""
    # (a) Unit contract for every keyed adapter.
    keyed = [("RAMP_DETAIL", compare_env.RAMP_DETAIL, "PM"),
             ("HIGHWAY_SEQUENCE", compare_env.HIGHWAY_SEQUENCE, "PM"),
             ("INTERSECTION_DETAIL", compare_env.INTERSECTION_DETAIL, "Post Mile"),
             ("HIGHWAY_DETAIL", compare_env.HIGHWAY_DETAIL, "Post Mile")]
    for name, adapter, key in keyed:
        assert adapter.key_col == key, (name, adapter.key_col)
        assert adapter._resolve_key_field(["County", key, "Desc"]) == 1, (name, "present")
        # case/whitespace tolerant when present
        assert adapter._resolve_key_field(["County", f"  {key.upper()} ", "X"]) == 1, \
            (name, "case/whitespace")
        # absent -> fail-closed raise (was a silent return 0)
        try:
            adapter._resolve_key_field(["County", "Desc"])
            assert False, (name, "a missing configured key column must raise")
        except ValueError as e:
            assert key in str(e) and adapter.REPORT_NAME in str(e), (name, str(e))
    # An unkeyed adapter (key_col=None) legitimately uses the first column.
    assert compare_env.HIGHWAY_LOG.key_col is None, compare_env.HIGHWAY_LOG.key_col
    assert compare_env.HIGHWAY_LOG._resolve_key_field(["A", "B", "C"]) == 0

    # (b) End-to-end: two IDENTICAL malformed Ramp Detail folders whose header
    #     lacks PM must fail closed (an error), never a clean match. Since
    #     CMP-AUD-032 pins the schema, a PM-less header is now refused as an
    #     UNRECOGNIZED layout (which runs before the key-column check) — a
    #     stricter fail-closed, still no false match and no workbook written.
    bad_header = [h for h in HEADER if h != "PM"]          # County, Ramp ID, Lighting
    bad_data = [[r[0]] + r[2:] for r in DATA_A]            # drop the PM cell
    root = Path(tempfile.mkdtemp())
    try:
        sheet = compare_env.RAMP_DETAIL.sheet_name
        a = root / "2026-06-16 ssor-prod" / "ramp_detail"
        b = root / "2026-06-16 ssor-test" / "ramp_detail"
        a.mkdir(parents=True)
        b.mkdir(parents=True)
        _write_route_file(a / f"ramp_detail_route_{ROUTE}.xlsx", sheet, bad_header, bad_data)
        _write_route_file(b / f"ramp_detail_route_{ROUTE}.xlsx", sheet, bad_header, bad_data)
        out = root / "cmp.xlsx"
        res = compare_env.RAMP_DETAIL.compare_folders(
            a.parent, b.parent, out, events=Events(),
            confirm_overwrite=lambda _p: True, mode="values")
        assert res.status == "error", ("must fail closed, not match", res.status)
        # Refused either as an unrecognized layout (032) or a missing key (028).
        msg = (res.message or "").lower()
        assert "recognized" in msg or "pm" in msg, ("names the refusal", res.message)
        assert not out.exists(), "no workbook may be written on a fail-closed layout"
    finally:
        shutil.rmtree(root, ignore_errors=True)


# --------------------------------------------------------------------------- #
# HF-04 — dual-layout cross-environment support (PCOA-FINAL-001)
# --------------------------------------------------------------------------- #
# getattr so a pre-fix tree runs these tests to SEMANTIC failures (the exact
# defect signatures) instead of crashing at import.
RD_HEADER_2026 = list(getattr(_rd, "_TSMIS_HEADER_2026",
                              ["Route", "Location", "PRE", "PM",
                               "Date of Record", "HG", "Area 4", "City Code",
                               "R/U", "OF", "TY", "Description"])[1:])


def _rd_row_valuepos(pm, desc, hg=""):
    """A classic-layout row at the export's true VALUE positions (census: the
    labels shift right of the values, so Description's VALUE sits at raw
    position 9 under the 'R/U' label and position 10 is empty). `_rd_row`
    above deliberately mirrors the LABEL positions for the positional
    same-layout tests; mixed-pair tests need the real shape, because the
    name-keyed projection reads the value positions the way CMP-AUD-046
    corrected them."""
    row = [""] * len(RD_HEADER)
    row[0] = "12-ORA-001"                       # Location
    row[1] = "R"                                # PM prefix value
    row[2] = pm                                 # PM value
    row[5] = hg                                 # HG value
    row[9] = desc                               # Description VALUE (under 'R/U')
    return row


def _rd_row_2026(pm, desc, of="F", ty="D", hg="D"):
    row = [""] * len(RD_HEADER_2026)
    row[0] = "12-ORA-001"
    row[RD_HEADER_2026.index("PRE")] = "R"
    row[RD_HEADER_2026.index("PM")] = pm
    row[RD_HEADER_2026.index("HG")] = hg
    row[RD_HEADER_2026.index("OF")] = of
    row[RD_HEADER_2026.index("TY")] = ty
    row[RD_HEADER_2026.index("Description")] = desc
    return row


def _run_env(root, data_a, header_a, data_b, header_b, mode="values"):
    sheet = compare_env.RAMP_DETAIL.sheet_name
    a = root / "2026-07-09 ssor-prod" / "ramp_detail"
    b = root / "2026-07-23 ssor-prod" / "ramp_detail"
    a.mkdir(parents=True)
    b.mkdir(parents=True)
    _write_route_file(a / f"ramp_detail_route_{ROUTE}.xlsx", sheet, header_a, data_a)
    _write_route_file(b / f"ramp_detail_route_{ROUTE}.xlsx", sheet, header_b, data_b)
    out = root / "cmp.xlsx"
    res = compare_env.RAMP_DETAIL.compare_folders(
        a.parent, b.parent, out, events=Events(),
        confirm_overwrite=lambda _p: True, mode=mode)
    return res, out


def test_new_layout_end_to_end():
    """A July-2026 pair compares with the export's own labels — OF/TY are
    LABELLED columns and compared like any other same-edition column."""
    root = Path(tempfile.mkdtemp())
    try:
        data_a = [_rd_row_2026("1.000", "001/ON-A"), _rd_row_2026("2.000", "001/OFF-B")]
        data_b = [_rd_row_2026("1.000", "001/ON-A"),
                  _rd_row_2026("2.000", "001/OFF-B", of="N")]   # a genuine OF change
        res, out = _run_env(root, data_a, RD_HEADER_2026, data_b, RD_HEADER_2026)
        assert res.status == "ok", (res.status, res.message)
        wb = load_workbook(out, read_only=True, data_only=True)
        body = list(wb["Comparison"].iter_rows(values_only=True))
        wb.close()
        header = [("" if c is None else str(c)) for c in body[0]]
        assert "OF" in header and "TY" in header and "PRE" in header, \
            ("the July-2026 pair must display the export's own labels", header)
        of_col = header.index("OF")
        rows = body[1:]
        marked = [r for r in rows
                  if isinstance(r[of_col], str) and _DIFF_MARK in r[of_col]]
        assert len(marked) == 1, \
            ("a genuine OF change on a same-edition pair must count", len(marked))
        assert res.comparison_outcome.counts.differing_cells == 1, \
            res.comparison_outcome.counts
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_mixed_layout_end_to_end():
    """HF-04's controlling decision: a MIXED classic/July-2026 pair compares
    correctly — the shared columns by NAME (a real Description change counts;
    identical data counts zero), the edition-specific columns as CONTEXT
    (never a diff, the exporting side's value shown), and the Notes sheet
    documents the projection."""
    root = Path(tempfile.mkdtemp())
    try:
        # Same underlying data in both editions + one REAL Description change
        # at PM 2.000; the July side carries OF/TY letters the classic cannot.
        data_old = [_rd_row_valuepos("1.000", "001/ON-A"),
                    _rd_row_valuepos("2.000", "001/OFF-B")]
        data_new = [_rd_row_2026("1.000", "001/ON-A", hg=""),
                    _rd_row_2026("2.000", "001/OFF-CHANGED", hg="")]
        res, out = _run_env(root, data_old, RD_HEADER, data_new, RD_HEADER_2026)
        assert res.status == "ok", (res.status, res.message)
        wb = load_workbook(out, read_only=True, data_only=True)
        sheets = wb.sheetnames
        body = list(wb["Comparison"].iter_rows(values_only=True))
        notes = ([str(r[0]) for r in wb["Notes"].iter_rows(values_only=True) if r and r[0]]
                 if "Notes" in sheets else [])
        wb.close()
        header = [("" if c is None else str(c)) for c in body[0]]
        for name in ("Location", "PR", "PM", "Date of Record", "HG", "Area 4",
                     "City Code", "R/U", "Description", "PM Suffix", "OF", "TY"):
            assert name in header, ("the mixed display header", name, header)
        rows = body[1:]
        desc_col = header.index("Description")
        of_col, ty_col = header.index("OF"), header.index("TY")
        marked = [(r[0], i) for r in rows for i, v in enumerate(r)
                  if isinstance(v, str) and _DIFF_MARK in v]
        assert len(marked) == 1 and marked[0][1] == desc_col, \
            ("exactly the one real Description change may count", marked)
        assert res.comparison_outcome.counts.differing_cells == 1, \
            res.comparison_outcome.counts
        # Context: OF/TY display the July side's letters, never a diff marker.
        # The key column shows the canonical route/county/postmile identity
        # display (CMP-AUD-045), decimal-canonical postmile included.
        by_pm = {str(r[header.index("PM")]): r for r in rows}
        k1 = "001 / ORA / 1"
        assert k1 in by_pm, ("canonical key display expected", sorted(by_pm))
        assert by_pm[k1][of_col] == "F" and by_pm[k1][ty_col] == "D", \
            ("context OF/TY must show the exporting side's value",
             by_pm[k1][of_col], by_pm[k1][ty_col])
        assert any("mixed export layouts" in n for n in notes), \
            ("the mixed pair documents itself in Notes", notes[:3])
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_mixed_identical_data_is_clean():
    """A mixed pair over identical shared data reports a clean MATCH — layout
    drift alone must produce zero differing cells and zero one-sided rows."""
    root = Path(tempfile.mkdtemp())
    try:
        data_old = [_rd_row_valuepos("1.000", "001/ON-A"),
                    _rd_row_valuepos("2.000", "001/OFF-B")]
        data_new = [_rd_row_2026("1.000", "001/ON-A", hg=""),
                    _rd_row_2026("2.000", "001/OFF-B", hg="")]
        res, _out = _run_env(root, data_old, RD_HEADER, data_new, RD_HEADER_2026)
        assert res.status == "ok", (res.status, res.message)
        counts = res.comparison_outcome.counts
        assert counts.differing_cells == 0 and counts.differing_rows == 0, counts
        assert counts.side_a_only_rows == 0 and counts.side_b_only_rows == 0, counts
        assert res.verdict == "match", res.verdict
    finally:
        shutil.rmtree(root, ignore_errors=True)


def test_unknown_layout_pairings_still_refuse():
    """Dual-layout acceptance must not widen into trust: a recognized side
    paired with an UNKNOWN layout still refuses (the merger vouches only for
    the two censused editions), and an unknown-vs-unknown pair still refuses."""
    unknown = ["Location", "PRE", "PM", "Date of Record", "HG", "Area 4",
               "City Code", "R/U", "OF", "TY", "XS", "Description"]
    row = ["12-ORA-001", "R", "1.000", "2026-01-01", "D", "Y", "C", "U", "F",
           "D", "x", "001/DESC"]
    for tag, header_a, data_a in (
            ("classic-vs-unknown", RD_HEADER, [_rd_row("1.000", "001/ON-A")]),
            ("unknown-vs-unknown", unknown, [list(row)])):
        root = Path(tempfile.mkdtemp())
        try:
            res, out = _run_env(root, data_a, header_a, [list(row)], unknown)
            assert res.status == "error", (tag, res.status, res.message)
            assert "recognized" in (res.message or ""), (tag, res.message)
            assert not out.exists(), (tag, "no workbook on a refused layout")
        finally:
            shutil.rmtree(root, ignore_errors=True)


def main():
    test_config_is_pm_keyed()
    test_pm_key_collapses_coarse_cascade()
    test_end_to_end_values_workbook()
    test_missing_key_column_fails_closed()
    test_new_layout_end_to_end()
    test_mixed_layout_end_to_end()
    test_mixed_identical_data_is_clean()
    test_unknown_layout_pairings_still_refuse()
    print("OK  COMPARE-RAMP-DETAIL-PM-KEY: Ramp Detail keys on PM; a mid-route "
          "ramp insert that cascades into 5 spurious diff cells under coarse "
          "keying collapses to ONE one-sided ramp / zero diff cells under PM "
          "keying, end to end through compare_folders.")


if __name__ == "__main__":
    main()
