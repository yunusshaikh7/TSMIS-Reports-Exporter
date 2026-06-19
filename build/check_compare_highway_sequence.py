"""Golden check for the cross-environment HIGHWAY SEQUENCE comparison
(scripts/compare_env.HIGHWAY_SEQUENCE + compare_core).

Closes the audit gap the ramp checks never covered: Highway Sequence cross-env
had no end-to-end adapter lock. Like Ramp Detail, Highway Sequence's first
column (County) is COARSE — it repeats for every location on a route — so the
original "key on the first column" behavior aligned rows POSITIONALLY within the
route, cascading a single mid-route insert into spurious field diffs. v0.11.0
set CompareSchema.key_field to the granular postmile ("PM") column
(compare_env.HIGHWAY_SEQUENCE.key_col="PM").

Highway Sequence is also the one cross-env report whose export carries REAL but
HEADER-LESS internal columns; compare_env._load_xlsx_side gives each a stable
"(col <Letter>)" label so it shows as a normal field instead of a blank column.
This check pins BOTH behaviors end to end through compare_folders.

check_compare_keyfield.py locks the generic key_field MECHANISM with a toy
schema; check_compare_ramp_detail.py locks the Ramp Detail adapter. This locks
the Highway Sequence adapter: that it keys on "PM", reads the per-route
"Highway Locations" sheet, labels unnamed columns "(col X)", and that a
mid-route insert produces ONE one-sided location with zero spurious diffs.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_highway_sequence.py
"""
import os
import shutil
import sys
import tempfile
from dataclasses import replace
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
from compare_core import _DIFF_MARK, count_diffs, keys_for, union_keys
from events import Events
from openpyxl import Workbook, load_workbook

# A miniature Highway Sequence route. County is the COARSE first column; PM is
# the granular postmile that identifies a location. Column 3 (Excel column C) is
# written with a BLANK header — a real-but-unnamed internal column, as the TSMIS
# export emits. Side B changes that unnamed column on one matched location AND
# inserts one new postmile mid-route.
ROUTE = "001"
HEADER_RAW = ["County", "PM", None, "Description"]      # col C header is blank
HEADER_LABELED = ["County", "PM", "(col C)", "Description"]
DATA_A = [
    ["LA", "1.000", "x", "ALPHA"],
    ["LA", "2.000", "y", "BETA"],
    ["LA", "3.000", "z", "GAMMA"],
]
DATA_B = [
    ["LA", "1.000", "x", "ALPHA"],
    ["LA", "2.000", "yy", "BETA"],          # unnamed (col C) changed: 1 diff cell
    ["LA", "2.500", "n", "NEW"],            # inserted mid-route: 1 one-sided
    ["LA", "3.000", "z", "GAMMA"],
]
ROWS_A = [[ROUTE] + r for r in DATA_A]
ROWS_B = [[ROUTE] + r for r in DATA_B]


def test_config_is_pm_keyed():
    hs = compare_env.HIGHWAY_SEQUENCE
    assert hs.key_col == "PM", ("Highway Sequence must key on PM", hs.key_col)
    assert hs.sheet_name == "Highway Locations", hs.sheet_name
    assert hs.subdir == "highway_sequence", hs.subdir


def test_pm_key_collapses_coarse_cascade():
    """Through the REAL schema the adapter builds (with the (col X)-labeled
    header): coarse County keying cascades the insert into spurious diffs; PM
    keying isolates it to one one-sided location plus the single real cell diff."""
    sc_pm = compare_env.HIGHWAY_SEQUENCE._schema(HEADER_LABELED, "SSOR-PROD", "SSOR-DEV")
    assert sc_pm.key_field == HEADER_LABELED.index("PM"), \
        ("the adapter must resolve PM to its header index", sc_pm.key_field)
    sc_coarse = replace(sc_pm, key_field=0)

    def counts(sc, kf):
        kt = keys_for(ROWS_A, True, kf)
        kn = keys_for(ROWS_B, True, kf)
        return count_diffs(sc, ROWS_A, ROWS_B, kt, kn, union_keys(kt, kn), True)

    coarse = counts(sc_coarse, 0)
    # County keying mis-pairs locations after the insert → spurious diffs.
    assert coarse["both"] == 3 and coarse["n_only"] == 1, coarse
    assert coarse["diff_cells"] > 1, \
        ("coarse County key must cascade the mid-route insert", coarse)

    pm = counts(sc_pm, sc_pm.key_field)
    # PM keying: 3 matched locations + 1 inserted one-sided; the only real diff
    # is the unnamed (col C) change on PM 2.000.
    assert pm["both"] == 3 and pm["t_only"] == 0 and pm["n_only"] == 1, pm
    assert pm["diff_cells"] == 1 and pm["diff_rows"] == 1, \
        ("PM key must isolate the insert; one real (col C) diff remains", pm)


def _write_route_file(path, sheet, header, data):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for row in data:
        ws.append(row)
    wb.save(path)
    wb.close()


def test_end_to_end_values_workbook():
    """Drive the actual compare_folders path over real per-route XLSX files
    (with a blank-header internal column) and read the VALUES workbook back: the
    unnamed column is labeled '(col C)', the mid-route insert is one one-sided
    location, and exactly one cell differs (the (col C) change)."""
    root = Path(tempfile.mkdtemp())
    try:
        sheet = compare_env.HIGHWAY_SEQUENCE.sheet_name
        a = root / "2026-06-16 ssor-prod" / "highway_sequence"
        b = root / "2026-06-16 ssor-dev" / "highway_sequence"
        a.mkdir(parents=True)
        b.mkdir(parents=True)
        _write_route_file(a / f"highway_sequence_route_{ROUTE}.xlsx", sheet,
                          HEADER_RAW, DATA_A)
        _write_route_file(b / f"highway_sequence_route_{ROUTE}.xlsx", sheet,
                          HEADER_RAW, DATA_B)

        out = root / "cmp.xlsx"
        res = compare_env.HIGHWAY_SEQUENCE.compare_folders(
            a.parent, b.parent, out, events=Events(),
            confirm_overwrite=lambda _p: True, mode="values")
        assert res.status == "ok", (res.status, res.message)
        assert res.verdict == "diff", res.verdict
        assert "DIFFERENCES FOUND" in res.summary_lines[0], res.summary_lines[0]

        wb = load_workbook(out, read_only=True, data_only=True)
        cmp_rows = list(wb["Comparison"].iter_rows(values_only=True))
        header = list(cmp_rows[0])
        body = cmp_rows[1:]
        wb.close()
        # The unnamed internal column shows as a labeled field, not a blank.
        assert "(col C)" in header, ("unnamed column must be labeled (col X)", header)
        # has_route layout: A=Route B=PM C=# D=A Row E=B Row F=Status G=Diffs H..
        statuses = [r[5] for r in body]
        one_sided = [s for s in statuses if s and s != "Both"]
        assert len(body) == 4, ("union = 3 matched + 1 inserted", len(body))
        assert statuses.count("Both") == 3, statuses
        assert len(one_sided) == 1 and one_sided[0].endswith("only"), one_sided
        neq = sum(1 for r in body for v in r
                  if isinstance(v, str) and _DIFF_MARK in v)
        assert neq == 1, ("exactly the one unnamed-column change differs", neq)
    finally:
        shutil.rmtree(root, ignore_errors=True)


def main():
    test_config_is_pm_keyed()
    test_pm_key_collapses_coarse_cascade()
    test_end_to_end_values_workbook()
    print("OK  COMPARE-HIGHWAY-SEQUENCE: keys on PM (coarse County cascade "
          "collapses to one one-sided location), reads the 'Highway Locations' "
          "sheet, labels unnamed internal columns '(col X)', and reports exactly "
          "the one real cell difference, end to end through compare_folders.")


if __name__ == "__main__":
    main()
