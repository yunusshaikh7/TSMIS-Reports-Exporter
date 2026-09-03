"""Golden check for the TSMIS-vs-TSN Highway Sequence comparator
(scripts/compare_highway_sequence_tsn.py) — the FLAT recipe with a COUNTY+PM key.

Locks: the CompareSchema wiring (PM key; key_normalizer = county|PM; context_fields
= HG + City + Distance To Next Point; the Notes legend_writer); the normalizers
(county trailing-period strip, prefix/PM/suffix re-glue, description route-prefix
strip + whitespace collapse); the TSN PDF parser's pure helpers (x-window bucketing,
2-char flag split, route/location regex); and end-to-end that

  * the same county+PM keys together even when TSMIS writes "LA." and TSN "LA",
  * the SAME postmile in two DIFFERENT counties is NOT confused (county-relative),
  * prefix/suffix glue ("R" + "010.179" == "R010.179"; "050.025" + "E" == "050.025E"),
  * FT and Description (after the "<route>/" strip) are GENUINE diffs,
  * HG / City / Distance contribute ZERO diff cells (context), and
  * a postmile only one side lists is a one-sided row.

No PDFs / real files; CI-safe. Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_highway_sequence_tsn.py
"""
import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_highway_sequence_tsn as hs
import consolidate_tsn_highway_sequence as ths
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []
DIFF = " ≠ "


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _write_tsmis(path, rows):
    """Synthetic CONSOLIDATED Highway Sequence: header[0]='Route'; loader reads by
    POSITION. rows = (route, county, city, prefix, pm, suffix, hg, ft, dist, desc)."""
    wb = Workbook()
    ws = wb.active
    ws.title = hs.TSMIS_SHEET
    ws.append(["Route", "County", "City", "", "PM", "", "HG", "FT",
               "Distance To Next Point", "Description"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    wb.close()


def _write_tsn(path, rows, version=None):
    """Synthetic NORMALIZED TSN workbook: ths.NORMALIZED_HEADER, read positionally,
    plus the v4 marker sheet (version=None writes the current one; 0 omits it —
    the pre-v4 shape). rows = (route, county, pm, city, hg, ft, dist, desc)."""
    wb = Workbook()
    ws = wb.active
    ws.title = ths.NORMALIZED_SHEET
    ws.append(ths.NORMALIZED_HEADER)
    for r in rows:
        ws.append(list(r))
    if version != 0:
        mk = wb.create_sheet(ths.MARKER_SHEET)
        mk.append(["Report", ths.REPORT_NAME])
        mk.append(["Normalization version",
                   ths.NORMALIZATION_VERSION if version is None else version])
    wb.save(path)
    wb.close()


def _comparison(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Comparison"]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
        return header, rows, wb.sheetnames
    finally:
        wb.close()


def test_schema():
    print("schema + normalizers:")
    sc = hs._SCHEMA
    check("key is PM", sc.header[sc.key_field] == "PM")
    check("side names TSMIS / TSN", sc.side_a == "TSMIS" and sc.side_b == "TSN")
    check("no scalar key_normalizer (typed PhysicalKey rows own identity)",
          sc.key_normalizer is None)
    # Owner decision 2026-08-10: HG / City / Distance To Next Point are COMPARED.
    # They were context because each is noisy for a structural reason (TSMIS
    # leaves HG blank for whole counties; TSN assigns city codes more
    # aggressively; distance is measured to each system's own next listed
    # point). The owner wants them counted anyway — the noise is the answer he
    # is asking for. Nothing else about HSL comparison changed.
    check("no context fields — every HSL column is compared",
          tuple(sc.context_fields) == ())
    check("Notes legend_writer set (the indicator)", sc.legend_writer is not None)
    check("county 'LA.' -> 'LA' (period strip)", hs._norm_county("LA.") == "LA")
    check("county 'MEN' -> 'MEN' (no-op)", hs._norm_county("MEN") == "MEN")
    check("glue prefix+PM ('R' + '010.179' -> 'R010.179')",
          hs._glue_pm("R", "010.179", None) == "R010.179")
    check("glue PM+suffix ('050.025' + 'E' -> '050.025E')",
          hs._glue_pm(None, "050.025", "E") == "050.025E")
    # CMP-AUD-204: the TSMIS side strips ONLY its own-route leading label;
    # the TSN side is verbatim (numeric prefixes are source claims).
    check("TSMIS desc strips its OWN-route label",
          hs._desc_tsmis("001/NB OFF TO X", "001") == "NB OFF TO X")
    check("TSMIS desc keeps a CROSS-route label",
          hs._desc_tsmis("1/103 SEP 53-145", "680") == "1/103 SEP 53-145")
    check("TSMIS own-route strip is padding-insensitive ('14/' on route 014)",
          hs._desc_tsmis("14/SOME RD", "014") == "SOME RD")
    check("TSN desc is verbatim (numeric prefix preserved)",
          hs._desc_plain("1/103 SEP 53-145") == "1/103 SEP 53-145")
    check("desc collapses double spaces",
          hs._desc_plain("SB ON  ARGYLE AV") == "SB ON ARGYLE AV")
    # CMP-AUD-197 (HSL half): the TSMIS Excel export's OOXML control escapes
    # are encoded CRs — _v decodes them exactly as the Stage-8 oracle's
    # xlsx-unescape does (byte-equivalent seam pinned in
    # check_compare_tsn_common), so the four censused `_x000d_` cells compare
    # as their real line-break content on the vs-TSN legs too.
    check("_v decodes the censused _x000d_ escape to a break-space",
          hs._v("WEIGH STA_x000d_(BOTH DIRS)") == "WEIGH STA (BOTH DIRS)")
    check("_v keeps the _x005F_-escaped literal token (OOXML spec)",
          hs._v("TAG_x005F_x000d_") == "TAG_x000d_")
    check("_desc_plain rides the decode (collapse after unescape)",
          hs._desc_plain("ABC_x000d_ DEF") == "ABC DEF")
    # CMP-AUD-045: the projector bakes the typed key — canonical identity is
    # the COMPLETE GLUED postmile (prefix + padded PM + equate suffix).
    row = hs._tsmis_row(
        ["001", "LA.", "DAPT", "R", "000.500", "E", "D", "H", "000.100", "JCT 5"])
    key = row[1 + hs.KEY_FIELD]
    comp = dict(key.physical_identity.canonical_components)
    check("typed key canonical = (001, LA, R000.500E)",
          comp == {"route": "001", "county": "LA", "postmile": "R000.500E"})
    check("typed key display payload is the glued PM", str(key) == "R000.500E")
    # the normalizer version pair: the workbook marker must match the catalog's
    # TsnEntry (D2 auto-rebuild keys off the catalog value).
    import report_catalog
    entry = next(t for t in report_catalog.TSN if t.subdir == "highway_sequence")
    check("catalog normalization_version == converter NORMALIZATION_VERSION",
          entry.normalization_version == ths.NORMALIZATION_VERSION == 4)


def test_parser_helpers():
    print("TSN PDF parser helpers (no PDF):")
    check("location regex accepts 'R010.179' / '050.025E' / '000.000'",
          all(ths.LOCATION_RE.match(x) for x in ("R010.179", "050.025E", "000.000")))
    check("location regex rejects 'POSTMILE' / 'EQUATES'",
          not ths.LOCATION_RE.match("POSTMILE") and not ths.LOCATION_RE.match("EQUATES"))
    check("county regex accepts 'MEN' / 'LA.' , rejects 'OTM22025'",
          ths.COUNTY_RE.match("MEN") and ths.COUNTY_RE.match("LA.")
          and not ths.COUNTY_RE.match("OTM22025"))
    check("route norm '5' -> '005', '5S' -> '005S'",
          ths._norm_route("5") == "005" and ths._norm_route("5S") == "005S")
    # x-window bucketing: county / postmile / flag / distance / description bands
    check("bucket county x=15", ths._bucket(15) == "county")
    check("bucket postmile x=105", ths._bucket(105) == "pm")
    check("bucket flag x=185", ths._bucket(185) == "flag")
    check("bucket distance x=226", ths._bucket(226) == "dist")
    check("bucket description x=300", ths._bucket(300) == "desc")
    # flag split: 'DH' -> HG 'D', FT 'H'
    flag = "DH"
    check("flag split -> HG=D, FT=H", flag[0] == "D" and flag[1] == "H")


def _sheet_text(path, sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return ""
        return "\n".join(
            " ".join("" if c is None else str(c) for c in r)
            for r in wb[sheet].iter_rows(values_only=True))
    finally:
        wb.close()


def test_end_to_end():
    print("end-to-end (typed county+PM identity, county-relative, glue, "
          "context non-asserting):")
    root = Path(tempfile.mkdtemp(prefix="tsmis_hs_tsn_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # (route, county, city, prefix, pm, suffix, hg, ft, dist, desc)
    _write_tsmis(tsmis, [
        # LA.000.500: county period differs (LA. vs LA); HG/City/Distance differ
        #   (context); description carries the "001/" prefix; FT equal -> identical row.
        ("001", "LA.", "LACITY", None, "000.500", None, "D", "H", "000.100", "001/JCT 5"),
        # LA.001.000: a GENUINE FT diff (H vs I).
        ("001", "LA.", None, None, "001.000", None, "D", "H", "000.200", "PT A"),
        # MEN.000.000 vs ORA.000.000: same PM, different county — must NOT confuse.
        ("001", "MEN", None, None, "000.000", None, None, "H", "000.056", "BEGIN MEN"),
        ("001", "ORA", None, "R", "000.129", None, "D", "H", "000.000", "BEGIN ORA"),
        # equate suffix glue: TSMIS prefix+PM+suffix -> '050.025E'
        ("001", "MEN", None, None, "050.025", "E", None, "H", "000.076", "PT E"),
        # a TSMIS-only postmile (TSN doesn't list it) -> one-sided.
        ("001", "LA", None, None, "002.000", None, "D", "H", None, "TSMIS ONLY PT"),
        # CMP-AUD-204 (the 81-false-cleans class): TSMIS prints the bare text
        # where TSN's verbatim value keeps its authoritative numeric prefix —
        # a REAL difference the old symmetric strip silently erased.
        ("001", "LA", None, None, "004.000", None, "D", "H", None, "103 SEP 53-145"),
    ])
    # (route, county, pm, city, hg, ft, dist, desc)
    _write_tsn(tsn, [
        ("001", "LA", "000.500", None, "U", "H", "000.050", "JCT 5"),   # FT eq, ctx differ
        ("001", "LA", "001.000", None, "U", "I", "000.090", "PT A"),    # FT diff H vs I
        ("001", "MEN", "000.000", None, "U", "H", "000.056", "BEGIN MEN"),
        ("001", "ORA", "R000.129", None, "D", "H", "000.000", "BEGIN ORA"),
        ("001", "MEN", "050.025E", None, "U", "H", "000.076", "PT E"),
        ("001", "LA", "003.000", None, "U", "H", "000.010", "TSN ONLY PT"),  # one-sided
        # CMP-AUD-204: TSN keeps its numeric prefix verbatim — a REAL diff vs
        # the TSMIS bare text (the old symmetric strip false-cleaned this).
        ("001", "LA", "004.000", None, "D", "H", None, "1/103 SEP 53-145"),
        # CMP-AUD-156: a pointer distance token rides the context column verbatim.
        ("001", "LA", "005.000", None, "U", "H", "*P*", "POINTER PT"),
        # CMP-AUD-158: a pre-county equate annotation (blank County) stays a
        # disclosed one-sided TSN row under the reserved marker.
        ("001", None, "R000.000", None, None, None, None, "EQUATES TO"),
    ])
    res = hs.compare(tsmis, tsn, out, events=Events(),
                     confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows, sheets = _comparison(out)
    check("Notes sheet present (the indicator)", "Notes" in sheets)

    # The key column ("PM") shows the typed identity's canonical display
    # "route / county / glued postmile" (CMP-AUD-045).
    pmcol = header.index("PM")
    cocol = header.index("County")
    by = {r[pmcol]: r for r in rows}

    ft = header.index("FT")
    desc = header.index("Description")
    k500 = "001 / LA / 000.500"
    check("LA 000.500 (period-normalized) keyed together — present once",
          k500 in by)
    check("LA 000.500 FT equal (H=H) — no diff", DIFF not in by[k500][ft])
    check("LA 000.500 Description '001/JCT 5' vs 'JCT 5' own-route strip — no diff",
          DIFF not in by[k500][desc])
    check("LA 000.500 County column shows normalized 'LA'", by[k500][cocol] == "LA")
    check("LA 001.000 FT H vs I is a GENUINE diff",
          DIFF in by["001 / LA / 001.000"][ft])
    check("MEN 000.000 and ORA R000.129 both present (county-relative, not merged)",
          "001 / MEN / 000.000" in by and "001 / ORA / R000.129" in by)
    check("MEN 000.000 not confused with ORA — its desc is BEGIN MEN",
          "BEGIN MEN" in by["001 / MEN / 000.000"][desc])
    check("equate suffix glue: MEN 050.025E keyed together (no diff on FT)",
          "001 / MEN / 050.025E" in by
          and DIFF not in by["001 / MEN / 050.025E"][ft])
    check("cross-route TSMIS token is a REAL Description diff (CMP-AUD-204)",
          DIFF in by["001 / LA / 004.000"][desc]
          and "1/103 SEP 53-145" in by["001 / LA / 004.000"][desc])

    # HG / City / Distance are COMPARED since 2026-08-10 — the fixture's HG and
    # City genuinely differ, so they must now be MARKED, not silently passed.
    ctx_cols = [header.index(c) for c in ("HG", "City", "Distance To Next Point")]
    ctx_diffs = sum(1 for r in rows for i in ctx_cols if DIFF in r[i])
    check("HG/City/Distance now carry diff markers (no longer context)",
          ctx_diffs > 0)

    # one-sided rows: TSMIS 002.000 / TSN 003.000, the pointer row, and the
    # blank-county annotation under its reserved marker (CMP-AUD-156/158)
    check("Only in TSMIS + Only in TSN sheets present",
          "Only in TSMIS" in sheets and "Only in TSN" in sheets)
    only_tsn = _sheet_text(out, "Only in TSN")
    check("pointer token '*P*' conserved verbatim in the one-sided TSN row",
          "*P*" in only_tsn)
    check("blank-county TSN annotation disclosed (R000.000, EQUATES TO)",
          "R000.000" in only_tsn and "EQUATES TO" in only_tsn)

    # a pre-v4 normalized workbook is refused with a rebuild hint
    old = root / "old.xlsx"
    _write_tsn(old, [("001", "LA", "000.500", None, "U", "H", "000.050", "X")],
               version=0)
    res_old = hs.compare(tsmis, old, root / "c2.xlsx", events=Events(),
                         confirm_overwrite=lambda _p: True, mode="values")
    check("pre-v4 TSN workbook refused with a rebuild hint",
          res_old.status == "error"
          and "older TSN Highway Sequence converter" in res_old.message)

    total = sum(1 for r in rows for c in r if DIFF in c)
    print(f"      (rows={len(rows)}, total diff cells={total}, context diff cells={ctx_diffs})")


def test_equate_seat():
    """Roadmap E10: the Excel export seats the equate "E" on the realignment
    record about a quarter of the time, where TSN puts it on the equated
    postmile's own row; with the complete glued postmile as identity BOTH rows
    of such a relation went one-sided. TSN now declares the relation and the
    export's suffix is seated on the target BEFORE the keys are built — through
    the shipped compare(), read back from the written workbook."""
    print("equate seat (E10): TSN declares, the export's suffix moves to the target:")
    root = Path(tempfile.mkdtemp(prefix="tsmis_hs_seat_"))
    tsmis, tsn, out = root / "t.xlsx", root / "n.xlsx", root / "c.xlsx"
    # (route, county, city, prefix, pm, suffix, hg, ft, dist, desc)
    _write_tsmis(tsmis, [
        # A — the class: the export seats the E on the realignment record.
        ("001", "CC", None, "R", "004.629", "E", "D", "H", None, "PM EQUATION"),
        ("001", "CC", None, "R", "004.508", None, "D", "H", "000.660", "BEGIN R REALIGNMENT"),
        # B — already seated on the target: untouched.
        ("001", "CC", None, None, "006.000", None, "D", "H", None, "PM EQUATION"),
        ("001", "CC", None, None, "005.900", "E", "D", "H", "000.100", "JCT ST 4"),
        # C — the export carries no E anywhere: an honest difference stays.
        ("001", "CC", None, None, "010.000", None, "D", "H", None, "PM EQUATION"),
        ("001", "CC", None, None, "009.800", None, "D", "H", "000.200", "END R REALIGNMENT"),
        # D — the target postmile lists TWO export rows and neither prints
        #     TSN's Description: ambiguous, left exactly as exported.
        ("001", "CC", None, None, "020.000", "E", "D", "H", None, "PM EQUATION"),
        ("001", "CC", None, None, "019.700", None, "D", "H", "000.000", "ROUTE RESUME"),
        ("001", "CC", None, None, "019.700", None, "D", "H", "000.300", "BRKINRTE"),
        # E — two export rows at the target; the one printing TSN's
        #     Description is the target and receives the E.
        ("001", "CC", None, None, "030.000", "E", "D", "H", None, "PM EQUATION"),
        ("001", "CC", None, None, "029.500", None, "D", "H", "000.000", "ROUTE RESUME"),
        ("001", "CC", None, None, "029.500", None, "D", "H", "000.199", "001/N JCT ST FAI 280"),
    ])
    # (route, county, pm, city, hg, ft, dist, desc)
    _write_tsn(tsn, [
        ("001", "CC", "R004.629", None, None, None, None, "EQUATES TO"),
        ("001", "CC", "R004.508E", None, "D", "H", "000.660", "BEGIN R REALIGNMENT"),
        ("001", "CC", "006.000", None, None, None, None, "EQUATES TO"),
        ("001", "CC", "005.900E", None, "D", "H", "000.100", "JCT ST 4"),
        ("001", "CC", "010.000", None, None, None, None, "EQUATES TO"),
        ("001", "CC", "009.800E", None, "D", "H", "000.200", "END R REALIGNMENT"),
        ("001", "CC", "020.000", None, None, None, None, "EQUATES TO"),
        ("001", "CC", "019.700E", None, "D", "H", "000.300", "VIA ANGELS CREST HWY"),
        ("001", "CC", "030.000", None, None, None, None, "EQUATES TO"),
        ("001", "CC", "029.500", None, "D", "H", "000.000", "ROUTE RESUME"),
        ("001", "CC", "029.500E", None, "D", "H", "000.199", "N JCT ST FAI 280"),
        # CMP-AUD-158: a pre-county annotation declares nothing the export can carry.
        ("001", None, "R000.000", None, None, None, None, "EQUATES TO"),
        ("001", "CC", "R000.100E", None, "D", "H", "000.010", "PRE-COUNTY TARGET"),
    ])
    rows_n, _ = hs._load_tsn(tsn)
    rel = hs.equate_relations(rows_n)
    check("TSN declares six relations (the blank-county one included)", len(rel) == 6)
    check("a relation names (route, ann county, ann bare PM, target county, "
          "target bare PM, suffix, target Description)",
          bool(rel) and rel[0] == ("001", "CC", "R004.629", "CC", "R004.508", "E",
                                   "BEGIN R REALIGNMENT"))
    check("the pre-county annotation carries its blank county",
          bool(rel) and rel[-1][1] == "" and rel[-1][4] == "R000.100")

    stats = {}
    rows_t, _rows_n, _ = hs._load_pair(tsmis, tsn, stats)
    keys = {str(r[1 + hs.KEY_FIELD]) for r in rows_t}
    check("A: the export's E moved to the target (R004.629 bare, R004.508E)",
          "R004.629" in keys and "R004.508E" in keys and "R004.629E" not in keys)
    check("B: an already-seated relation is untouched",
          "006.000" in keys and "005.900E" in keys)
    check("C: an export carrying no E anywhere gets none invented",
          "010.000" in keys and "009.800" in keys and "009.800E" not in keys)
    check("D: an ambiguous target group is left exactly as exported",
          "020.000E" in keys and "019.700E" not in keys)
    check("E: the row printing TSN's Description is the target",
          "030.000" in keys and "029.500E" in keys)
    seated_e = [r for r in rows_t if str(r[1 + hs.KEY_FIELD]) == "029.500E"]
    check("...and it is the JCT row, not the ROUTE RESUME row",
          len(seated_e) == 1 and seated_e[0][-1] == "N JCT ST FAI 280")
    expected = {"declared": 6, "seated": 2, "ambiguous": 1,
                "no suffix at the realignment record": 2,
                "realignment record not exported": 1}
    check("stats: 6 declared, 2 seated, 1 ambiguous, 2 with no suffix at the record, "
          "1 record not exported (the pre-county annotation)", stats == expected)
    if stats != expected:
        print(f"      stats={stats}")
    plain, _ = hs._load_tsmis(tsmis)
    check("the plain loader (PDF flavors) still reads the export's own seat",
          "R004.629E" in {str(r[1 + hs.KEY_FIELD]) for r in plain})

    res = hs.compare(tsmis, tsn, out, events=Events(),
                     confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    header, rows, sheets = _comparison(out)
    by = {r[header.index("PM")]: r for r in rows}
    only_tsmis = _sheet_text(out, "Only in TSMIS")
    only_tsn = _sheet_text(out, "Only in TSN")
    check("A pairs on both rows in the written workbook",
          "001 / CC / R004.629" in by and "001 / CC / R004.508E" in by)
    check("...so neither A row is one-sided",
          "R004.629" not in only_tsmis and "R004.508" not in only_tsn)
    check("A's annotation still reports its by-design Description difference "
          "(PM EQUATION vs EQUATES TO)",
          DIFF in by["001 / CC / R004.629"][header.index("Description")])
    check("C stays honestly one-sided on the target (TSN 009.800E / TSMIS 009.800)",
          "009.800E" in only_tsn and "009.800" in only_tsmis)
    check("D stays as exported (TSMIS 020.000E and TSN 019.700E one-sided)",
          "020.000E" in only_tsmis and "019.700E" in only_tsn)
    check("the blank-county annotation is still disclosed one-sided",
          "R000.000" in only_tsn)
    summary = _sheet_text(out, "Summary")
    notes = _sheet_text(out, "Notes")
    check("the Summary discloses the seat counts",
          "Equate suffix seat: TSN declares 6" in summary and "for 2 relations" in summary
          and "1 ambiguous" in summary)
    check("...and the Notes carry the rule and the same counts",
          "moved to that target row before comparing" in notes
          and "TSN declares 6" in notes)
    shutil.rmtree(root, ignore_errors=True)


def main():
    test_schema()
    test_parser_helpers()
    test_end_to_end()
    test_equate_seat()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-HIGHWAY-SEQUENCE-TSN CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
