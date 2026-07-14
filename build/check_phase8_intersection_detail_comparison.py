#!/usr/bin/env python3
"""Permanent adversarial gate for the Stage-8 Intersection Detail oracle."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
import json
import sys
import tempfile

from openpyxl import Workbook


BUILD_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(BUILD_ROOT))
import phase8_intersection_detail_comparison as oracle  # noqa: E402
from phase3_independent_oracle import compare_rows  # noqa: E402


passed = 0


def require(condition: bool, message: str) -> None:
    global passed
    if not condition:
        raise AssertionError(message)
    passed += 1


def rejects(action, message: str) -> None:
    try:
        action()
    except (oracle.AuditError, ValueError, TypeError):
        require(True, message)
    else:
        raise AssertionError(message)


def row(index: int, *, county: str = "AAA", district: str = "01",
        pp: str = "", pm: str = "1.000", description: str = "X",
        source_only: tuple[tuple[str, object], ...] = ()) -> oracle.DetailRow:
    values = tuple(
        pp if field == "PR" else description if field == "Description" else ""
        for field in oracle.ASSERTED_FIELDS)
    return oracle.DetailRow(
        source_index=index, source="synthetic", source_ref=f"row {index}",
        route="001", district=district, county=county, complete_pp=pp,
        pm=pm, numeric_pm=oracle._numeric_pm(pm), route_suffix="",
        values=values, location_literal=f"{district} {county} 001",
        source_only=source_only)


def report_view_sheet(workbook: Workbook, *, claims: bool):
    worksheet = workbook.active
    worksheet.title = "Report View"
    headers = [[None] * 26 for _ in range(4)]
    headers[0][24] = "TSN only"
    headers[1][24] = "ML 2nd EFF"
    headers[1][25] = "ML ADT"
    headers[2][24] = "TSN only"
    headers[3][25] = "CS ADT"
    for header in headers:
        worksheet.append(header)
    main = [None] * 26
    cross = [None] * 26
    main[2] = cross[2] = "001"
    main[5] = "1.000"
    if claims:
        main[24], main[25], cross[25] = "2026-01-02", "100", "200"
    worksheet.append(main)
    worksheet.append(cross)
    return worksheet


def main() -> int:
    require(oracle._location("01 AAA 1") == ("01", "AAA", "001", ""),
            "valid Location normalization drift")
    require(oracle._location("12-ORA.-210U") == ("12", "ORA", "210", "U"),
            "suffix Location normalization drift")
    rejects(lambda: oracle._location("01 001"),
            "malformed Location was accepted")
    require(oracle._numeric_pm("001.000") == "1",
            "numeric PM canonicalization drift")
    require(oracle._numeric_pm("0.000") == "0",
            "zero PM canonicalization drift")
    rejects(lambda: oracle._numeric_pm("NOT-PM"),
            "nonnumeric PM was accepted")

    left = [row(0, county="AAA", description="ALPHA"),
            row(1, county="BBB", description="BETA")]
    right = [row(0, county="AAA", description="BETA"),
             row(1, county="BBB", description="ALPHA")]
    strong = oracle._comparison("county swap", left, right)
    require(strong["counts"]["differing_rows"] == 2
            and strong["counts"]["per_field_counts"] == {"Description": 2},
            "physical County identity masked a cross-county value swap")
    weak = compare_rows(
        oracle.PRODUCT_SCHEMA,
        tuple(item.product_oracle_row() for item in left),
        tuple(item.product_oracle_row() for item in right))
    require(weak.counts.differing_cells == 0,
            "weak Route+PM negative control no longer demonstrates masking")

    district = oracle._comparison(
        "district", [row(0, district="01")], [row(0, district="02")])
    require(district["counts"]["per_field_counts"] == {"District": 1},
            "District escaped the asserted source schema")
    prefix = oracle._comparison(
        "complete PP", [row(0, pp="L")], [row(0, pp="R")])
    require(prefix["counts"]["paired_rows"] == 0
            and prefix["counts"]["side_a_only_rows"] == 1
            and prefix["counts"]["side_b_only_rows"] == 1,
            "complete PP escaped physical identity")

    require(oracle._product_text(" A  B ") == "A B",
            "ordinary ASCII-space normalization drift")
    require(oracle._product_text("A\t") == "A\t",
            "tab source data was globally whitespace-folded")
    require(oracle._product_text(1.0) == "1",
            "integer-valued float normalization drift")

    counter_a = Counter({("b",): 1, ("a",): 2})
    counter_b = Counter({("a",): 2, ("b",): 1})
    require(oracle._counter_digest(counter_a) == oracle._counter_digest(counter_b),
            "counter digest depends on insertion order")
    require(oracle._counter_difference_examples(
        Counter({("x",): 1}), Counter({("y",): 1})) == {
            "missing_first_10": [{"entry": ["y"], "count": 1}],
            "extra_first_10": [{"entry": ["x"], "count": 1}]},
            "counter delta diagnostic drift")

    claims = (("MAIN_EFF_DATE", "2026-01-02"),
              ("MAIN_ADT", 100), ("CROSS_ADT", "200"))
    claim_row = row(0, source_only=claims)
    expected_claim = Counter({
        ("001", "1.000", "2026-01-02", "100", "200"): 1})
    require(oracle._source_only_counter([claim_row]) == expected_claim,
            "source-only claim mapping drift")
    raw_book = Workbook()
    raw_sheet = report_view_sheet(raw_book, claims=True)
    raw_result = oracle._inspect_report_view(
        raw_sheet, expected_records=1, raw_tsn=[claim_row])
    require(raw_result["source_only_mapping_exact"]
            and raw_result["source_only_nonblank_counts"] == {
                "MAIN_EFF_DATE": 1, "MAIN_ADT": 1, "CROSS_ADT": 1},
            "valid raw Report View mapping did not pass")
    raw_book.close()
    normalized_book = Workbook()
    normalized_sheet = report_view_sheet(normalized_book, claims=False)
    normalized_result = oracle._inspect_report_view(
        normalized_sheet, expected_records=1, raw_tsn=None)
    require(normalized_result["source_only_nonblank_counts"] == {
        "MAIN_EFF_DATE": 0, "MAIN_ADT": 0, "CROSS_ADT": 0},
        "normalized Report View blank contract drift")
    normalized_book.close()
    mutated_book = Workbook()
    mutated_sheet = report_view_sheet(mutated_book, claims=False)
    mutated_sheet.cell(5, 25).value = "2026-01-02"
    rejects(lambda: oracle._inspect_report_view(
        mutated_sheet, expected_records=1, raw_tsn=None),
        "unexpected normalized source-only value was accepted")
    mutated_book.close()

    raw_values = [None] * len(oracle.TSMIS_HEADERS)
    raw_values[0] = "R"
    raw_values[1] = "001.000"
    raw_values[2] = "U"
    raw_values[3] = "01 AAA 001U"
    raw_values[29] = ""
    truth = oracle._detail_row_from_tsmis(
        raw_values, source_index=0, source="synthetic TSMIS",
        source_ref="row 2", member_route="001U")
    with tempfile.TemporaryDirectory(prefix="id-stage8-gate-") as temporary:
        root = Path(temporary)
        consolidated = root / "consolidated.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Intersection Detail"
        worksheet.append(["Route", *oracle.TSMIS_HEADERS])
        serialized = list(raw_values)
        serialized[29] = None
        worksheet.append(["001U", *serialized])
        workbook.save(consolidated)
        workbook.close()
        inspected = oracle._inspect_consolidated(consolidated, [truth])
        require(inspected["projection_exact"]
                and inspected["explicit_member_route_mismatches"] == 0
                and inspected["explicit_physical_s_mismatches"] == 0
                and inspected["blank_string_to_physical_blank_cells"] == 1,
                "valid consolidated source projection did not pass")

        bad_route = root / "bad-route.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Intersection Detail"
        worksheet.append(["Route", *oracle.TSMIS_HEADERS])
        worksheet.append(["002", *raw_values])
        workbook.save(bad_route)
        workbook.close()
        rejects(lambda: oracle._inspect_consolidated(bad_route, [truth]),
                "mutated explicit member Route was accepted")

        bad_s = root / "bad-s.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Intersection Detail"
        worksheet.append(["Route", *oracle.TSMIS_HEADERS])
        changed = list(raw_values)
        changed[2] = "S"
        worksheet.append(["001U", *changed])
        workbook.save(bad_s)
        workbook.close()
        rejects(lambda: oracle._inspect_consolidated(bad_s, [truth]),
                "mutated explicit physical S was accepted")

        formula_path = root / "formulas.xlsx"
        values_path = root / "values.xlsx"
        workbook = Workbook()
        workbook.active["A1"] = "=1+1"
        workbook.save(formula_path)
        workbook.close()
        workbook = Workbook()
        workbook.active["A1"] = 2
        workbook.save(values_path)
        workbook.close()
        require(oracle._formula_tag_count(formula_path) == 1,
                "formula package tag count drift")
        require(oracle._formula_tag_count(values_path) == 0,
                "values package contains a formula tag")

        projection_book = Workbook()
        projection = projection_book.active
        projection.title = "__CMP_E2_SNAPSHOT_A"
        projection.append(["Source row", "Route", *oracle.SHARED_HEADER,
                           "Key (helper)"])
        projection.append([1, *claim_row.product_projection(),
                           "__CMP_E2_KEY_V1_00000001"])
        projection_result = oracle._inspect_projection_sheet(
            projection, [claim_row], side="TSN", snapshot=True)
        require(projection_result["projection_exact"],
                "valid snapshot projection did not pass")
        projection_book.close()

        only_book = Workbook()
        only = only_book.active
        only.title = "Only in A"
        only.append(["Route", "PM", "#", "A Row", "Missing from B",
                     *oracle.ASSERTED_FIELDS])
        only.append(["001", "1.000", 1, 2, None, *claim_row.values])
        expected_only = Counter({(
            "001", "1.000",
            *(oracle._product_text(value) for value in claim_row.values)): 1})
        only_result = oracle._inspect_only_sheet(
            only, present="A", missing="B", expected=expected_only)
        require(only_result["inventory_exact"],
                "valid one-sided inventory did not pass")
        only_book.close()

    steps = [0, *(46 * index for index in range(1, 21)), 982]
    edges_a = tuple(27.75 + 0.75 * step for step in steps)
    edges_b = (*edges_a[:4], *edges_a[7:])
    observed_a, observed_b = oracle._validate_pdf_profiles(
        Path("synthetic.pdf"), {21: Counter({edges_a: 1}),
                                18: Counter({edges_b: 1})})
    require(observed_a == edges_a and observed_b == edges_b,
            "valid per-document PDF grid was rejected")
    off_lattice = list(edges_a)
    off_lattice[5] += 0.1
    rejects(lambda: oracle._validate_pdf_profiles(
        Path("mutated.pdf"), {21: Counter({tuple(off_lattice): 1}),
                              18: Counter({edges_b: 1})}),
        "off-lattice PDF grid was accepted")
    a = [""] * 21
    b = [""] * 18
    a[20] = "RESIDUE"
    rejects(lambda: oracle._pdf_row(a, b),
            "nonblank vestigial PDF cell was accepted")

    require(set(oracle.EXPECTED_PRODUCT_COUNTS) == {
        "excel_vs_tsn_raw", "excel_vs_tsn_normalized", "pdf_vs_tsn_raw",
        "pdf_vs_tsn_normalized", "pdf_vs_excel"},
        "frozen five-leg product universe drift")
    require(oracle.EXPECTED_SOURCE_COUNTS["raw_vs_normalized"] == {
        "known": True, "paired_rows": 16626, "side_a_only_rows": 0,
        "side_b_only_rows": 0, "differing_rows": 0, "differing_cells": 0,
        "per_field_counts": {}, "asserted_cells": 565284,
        "context_cells": 0}, "raw-normalized frozen truth drift")

    print(json.dumps({"status": "pass", "assertions": passed},
                     separators=(",", ":")))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
