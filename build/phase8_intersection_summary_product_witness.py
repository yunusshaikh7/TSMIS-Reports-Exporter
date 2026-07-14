#!/usr/bin/env python3
"""Run the production Intersection Summary pipeline in an isolated destination.

This helper contains no oracle logic.  The independent Stage-8 driver invokes it
only after deriving source truth, then parses the emitted workbooks itself.  Product
imports stay in this child process so application parser/schema objects cannot become
accidental truth-side dependencies.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
import sys

from openpyxl import load_workbook


BUILD_ROOT = Path(__file__).resolve().parent
REPO_ROOT = BUILD_ROOT.parent
SCRIPTS_ROOT = REPO_ROOT / "scripts"


class QuietEvents:
    def on_log(self, _message: str) -> None:
        return None

    def is_cancelled(self) -> bool:
        return False


def _sha(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _loaded_product_manifest() -> dict[str, object]:
    entries = []
    root = SCRIPTS_ROOT.resolve()
    for name, module in sorted(sys.modules.items()):
        raw = getattr(module, "__file__", None)
        if not raw:
            continue
        path = Path(raw).resolve()
        try:
            relative = path.relative_to(root).as_posix()
        except ValueError:
            continue
        if path.suffix.lower() not in {".py", ".pyw"} or not path.is_file():
            continue
        entries.append({
            "module": name,
            "relative_path": relative,
            "bytes": path.stat().st_size,
            "sha256": _sha(path),
        })
    payload = json.dumps(
        entries, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return {
        "file_count": len(entries),
        "canonical_json_sha256": hashlib.sha256(payload).hexdigest(),
        "entries": entries,
    }


def _result_summary(result: object) -> dict[str, object]:
    comparison = getattr(result, "comparison_outcome", None)
    counts = getattr(comparison, "counts", None)
    count_payload = None
    if counts is not None:
        count_payload = {
            "known": counts.known,
            "paired_rows": counts.paired_rows,
            "side_a_only_rows": counts.side_a_only_rows,
            "side_b_only_rows": counts.side_b_only_rows,
            "differing_rows": counts.differing_rows,
            "differing_cells": counts.differing_cells,
            "asserted_cells": counts.asserted_cells,
            "context_cells": counts.context_cells,
            "per_field_counts": dict(counts.per_field_counts),
        }
    generation = getattr(result, "artifact_generation", None)
    generation_payload = None
    if generation is not None:
        generation_payload = {
            "completion": generation.completion,
            "publication_state": generation.publication_state,
            "requested_mode": generation.requested_mode,
            "members": [
                {
                    "flavor": member["flavor"],
                    "commit_role": member["commit_role"],
                    "path": member["path"],
                    "bytes": member["size"],
                    "sha256": member["sha256"],
                }
                for member in generation.members
            ],
        }
    return {
        "status": getattr(result, "status", None),
        "completion": getattr(result, "completion", None),
        "verdict": getattr(result, "verdict", None),
        "skipped_inputs": getattr(result, "skipped_inputs", None),
        "failed_inputs": getattr(result, "failed_inputs", None),
        "summary_lines": list(getattr(result, "summary_lines", ()) or ()),
        "counts": count_payload,
        "warnings": list(getattr(comparison, "warnings", ()) or ())
        if comparison else [],
        "failures": list(getattr(comparison, "failures", ()) or ())
        if comparison else [],
        "artifact_generation": generation_payload,
    }


def _mutated_copy(source: Path, destination: Path, mutate) -> Path:
    """Save one disposable workbook mutation below the isolated work root."""
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        mutate(workbook)
        workbook.save(destination)
    finally:
        workbook.close()
    return destination


def _known_gap_probes(compare_module, layout_module, consolidated: Path,
                      tsn_xlsx: Path, work_root: Path) -> dict[str, object]:
    """Measure known product-red acceptance paths without supplying oracle truth.

    These probes intentionally ask only whether the production loaders accept a
    malformed aggregate.  The Stage-8 driver independently decides what is correct.
    Keeping the probes in this child process prevents product schema/parser objects
    from leaking onto the truth side of the audit.
    """
    probe_root = work_root / "known-gap-probes"
    probe_root.mkdir(parents=True, exist_ok=True)
    baseline = compare_module._load_tsmis(consolidated)
    baseline_tsn = compare_module._load_tsn(tsn_xlsx)

    missing_total = _mutated_copy(
        consolidated, probe_root / "missing-total.xlsx",
        lambda wb: setattr(wb["Intersection Summary"]["B1"], "value", None))
    missing_total_counts = compare_module._load_tsmis(missing_total)

    def zero_categories(workbook) -> None:
        ws = workbook["Intersection Summary"]
        for row in range(2, ws.max_row + 1):
            for column in range(3, ws.max_column + 1):
                ws.cell(row, column).value = 0

    zeroed = _mutated_copy(
        consolidated, probe_root / "all-zero-categories.xlsx", zero_categories)
    zeroed_counts = compare_module._load_tsmis(zeroed)

    def loose_types(workbook) -> None:
        ws = workbook["Intersection Summary"]
        ws["C2"] = True
        ws["D2"] = 1.5

    loose = _mutated_copy(
        consolidated, probe_root / "loose-count-types.xlsx", loose_types)
    loose_counts = compare_module._load_tsmis(loose)

    def duplicate_tsn(workbook) -> None:
        ws = workbook[workbook.sheetnames[0]]
        ws.append([ws["A2"].value, 1])

    duplicate_tsn_path = _mutated_copy(
        tsn_xlsx, probe_root / "duplicate-tsn-category.xlsx", duplicate_tsn)
    duplicate_tsn_counts = compare_module._load_tsn(duplicate_tsn_path)

    def drop_route(workbook) -> None:
        ws = workbook["Intersection Summary"]
        ws.delete_rows(ws.max_row, 1)

    dropped = _mutated_copy(
        consolidated, probe_root / "dropped-route.xlsx", drop_route)
    dropped_counts = compare_module._load_tsmis(dropped)

    def duplicate_route(workbook) -> None:
        ws = workbook["Intersection Summary"]
        ws.append([ws.cell(2, column).value
                   for column in range(1, ws.max_column + 1)])

    duplicated = _mutated_copy(
        consolidated, probe_root / "duplicate-route.xlsx", duplicate_route)
    duplicated_counts = compare_module._load_tsmis(duplicated)

    ru_header = layout_module._IS_RURAL_URBAN
    control_header = layout_module._IS_CONTROL_TYPES
    spec = layout_module.INTERSECTION_SUMMARY_SPEC
    orphan = layout_module.counts_from_rows(
        spec, [(None, ru_header), (5, "-O OUTSIDE CITY")])
    countless_parent = layout_module.counts_from_rows(
        spec, [(None, ru_header), (None, "U-URBAN -I INSIDE CITY"),
               (5, "-O OUTSIDE CITY")])
    distinct_fold = layout_module.counts_from_rows(
        spec, [(None, control_header), (1, "J-SIGNAL PRETIMED"),
               (2, "P-SIGNALS FULL-TRAFFIC ACTUATED")])
    repeated_fold = layout_module.counts_from_rows(
        spec, [(None, control_header), (1, "J-SIGNAL PRETIMED"),
               (2, "J-SIGNAL PRETIMED")])

    total_slug = "total_intersections"
    typed_slugs = [slug for slug in baseline if slug != total_slug][:2]
    first_slug, second_slug = typed_slugs
    first_tsn_slug = next(iter(baseline_tsn))
    rural_outside_slug = "is_rural_urban_suburban_r_o"
    urban_outside_slug = "is_rural_urban_suburban_u_o"
    signal_slug = "is_control_types_s"
    results = {
        "missing_total_accepted_as_zero": (
            total_slug not in missing_total_counts
            and compare_module._rows(missing_total_counts, "tsmis")[-1][1] == 0),
        "all_zero_categories_with_nonzero_total_accepted": (
            zeroed_counts.get(total_slug) == baseline.get(total_slug)
            and all(value == 0 for slug, value in zeroed_counts.items()
                    if slug != total_slug)),
        "boolean_and_fraction_counts_coerced": (
            loose_counts[first_slug] == baseline[first_slug] + 1
            and loose_counts[second_slug] == baseline[second_slug] + 1),
        "duplicate_tsn_category_summed": (
            duplicate_tsn_counts[first_tsn_slug] == baseline_tsn[first_tsn_slug] + 1),
        "dropped_route_accepted_without_universe_diagnostic": (
            dropped_counts.get(total_slug, 0) < baseline.get(total_slug, 0)),
        "duplicate_route_accepted_and_double_counted": (
            duplicated_counts.get(total_slug, 0) > baseline.get(total_slug, 0)),
        "orphan_outside_city_defaults_to_rural": (
            orphan.get(rural_outside_slug) == 5
            and urban_outside_slug not in orphan),
        "countless_urban_parent_is_ignored": (
            countless_parent.get(rural_outside_slug) == 5
            and urban_outside_slug not in countless_parent),
        "distinct_j_p_fold_is_permitted": distinct_fold.get(signal_slug) == 3,
        "repeated_j_is_silently_summed": repeated_fold.get(signal_slug) == 3,
    }
    return {
        "all_expected_product_red_paths_reproduced": all(results.values()),
        "probes": results,
        "baseline_total": baseline.get(total_slug),
        "dropped_route_total": dropped_counts.get(total_slug),
        "duplicate_route_total": duplicated_counts.get(total_slug),
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-root", type=Path, required=True)
    parser.add_argument("--tsn-xlsx", type=Path, required=True)
    parser.add_argument("--work-root", type=Path, required=True)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    args.work_root.mkdir(parents=True, exist_ok=True)
    consolidated = args.work_root / "intersection_summary_consolidated.xlsx"
    comparison = args.work_root / "intersection_summary_comparison.xlsx"
    for path in (
        consolidated,
        comparison,
        comparison.with_name(f"{comparison.stem} (values){comparison.suffix}"),
    ):
        if path.exists():
            path.unlink()

    sys.path.insert(0, str(SCRIPTS_ROOT))
    import consolidate_intersection_summary  # type: ignore
    import compare_intersection_summary_tsn  # type: ignore
    import summary_layout  # type: ignore

    events = QuietEvents()
    consolidation = consolidate_intersection_summary.consolidate(
        events=events,
        confirm_overwrite=lambda _path: True,
        input_dir=args.xlsx_root,
        out_path=consolidated,
    )
    if consolidation.status != "ok":
        raise RuntimeError(f"production consolidation failed: {consolidation!r}")

    compared = compare_intersection_summary_tsn.compare(
        consolidated,
        args.tsn_xlsx,
        comparison,
        events=events,
        confirm_overwrite=lambda _path: True,
        mode="both",
    )
    if compared.status != "ok":
        raise RuntimeError(f"production comparison failed: {compared!r}")

    values = comparison.with_name(
        f"{comparison.stem} (values){comparison.suffix}")
    payload = {
        "consolidation": {
            "status": consolidation.status,
            "completion": consolidation.completion,
            "skipped_inputs": consolidation.skipped_inputs,
            "failed_inputs": consolidation.failed_inputs,
            "summary_lines": list(consolidation.summary_lines),
            "path": str(consolidated),
            "bytes": consolidated.stat().st_size,
            "sha256": _sha(consolidated),
        },
        "comparison": _result_summary(compared),
        "outputs": {
            "consolidated": str(consolidated),
            "formulas": str(comparison),
            "values": str(values),
        },
        "known_product_gap_probes": _known_gap_probes(
            compare_intersection_summary_tsn, summary_layout,
            consolidated, args.tsn_xlsx, args.work_root),
        "loaded_product_code": _loaded_product_manifest(),
    }
    encoded = (json.dumps(
        payload, ensure_ascii=False, separators=(",", ":")
    ) + "\n").encode("utf-8")
    sys.stdout.buffer.write(encoded)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
