"""CMP-AUD-037 — the DIRECT-path TSN normalization-marker gate (RD/ID/HD).

A classic (non-matrix) file comparison trusted ANY workbook that carried the
normalized sheet: a library built by an older normalizer was silently compared,
resurrecting whatever that version got wrong (an otherwise-identical row split
into two one-sided rows for "005" vs "5", "LA 000.500" vs "LA 0.5"). The
XLSX-sourced families (Ramp Detail / Intersection Detail / Highway Detail) now
stamp their normalized workbook with an in-workbook "TSN Normalization" marker,
and each direct loader refuses a pre-current file with a rebuild hint. The
matrix/library path already gated via the certificate (report_catalog's
normalization_version, D2); this closes the direct path.

This is the SHARED instrument. It locks:
  1. the compare_tsn_common helpers (write/read/require; write-only-workbook
     compatibility; absent -> 0; malformed -> 0 fail-safe),
  2. the MIRROR invariant — each comparator's NORMALIZATION_VERSION equals the
     catalog's normalization_version (drift would write a marker the gate then
     rejects, or accept a stale library, silently),
  3. the build_normalized writer seam actually stamps the requested version and
     leaves the data sheet intact (and writes NO marker when asked not to).
The per-family refusal/acceptance flows live in each family's own check.
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_tsn_common as ctc
import outcome
import tsn_library
from events import ConsolidateResult
from openpyxl import Workbook, load_workbook

import compare_highway_detail_tsn as hdt
import compare_highway_sequence_tsn as hsl
import compare_intersection_detail_tsn as idt
import compare_ramp_detail_tsn as rd
import tsn_load_highway_detail as _load_hd
import tsn_load_intersection_detail as _load_id
import tsn_load_ramp_detail as _load_rd

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def test_helper():
    print("shared marker helpers (compare_tsn_common):")
    wb = Workbook()
    ctc.write_normalization_marker(wb, 7, report_name="X")
    check("write -> read == 7 (normal wb)", ctc.normalization_marker_version(wb) == 7)
    check("marker sheet is named 'TSN Normalization'",
          ctc.NORMALIZATION_MARKER_SHEET in wb.sheetnames)

    # The real normalized workbook is write-only; ws['A1']= TypeErrors there, so
    # the helper must use create_sheet + append. Prove it round-trips through a save.
    wo = Workbook(write_only=True)
    ws = wo.create_sheet("Data")
    ws.append(["Route", "A"])
    ws.append(["001", "x"])
    ctc.write_normalization_marker(wo, 3)                 # must not raise on write-only
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "wo.xlsx"
        wo.save(p)
        rwb = load_workbook(p, read_only=True, data_only=True)
        try:
            check("write-only wb carries a readable marker == 3",
                  ctc.normalization_marker_version(rwb) == 3)
        finally:
            rwb.close()

    check("absent marker sheet -> 0", ctc.normalization_marker_version(Workbook()) == 0)
    bad = Workbook()
    bad.create_sheet(ctc.NORMALIZATION_MARKER_SHEET).append(
        ["Normalization version", "not-an-int"])
    check("malformed marker value -> 0 (fail-safe)",
          ctc.normalization_marker_version(bad) == 0)

    cur = Workbook()
    ctc.write_normalization_marker(cur, 5)
    raised = False
    try:
        ctc.require_current_normalization(cur, "x.xlsx", 6, "pre-v6: test")
    except ValueError as e:
        raised = "older TSN converter" in str(e) and "rebuild" in str(e)
    check("require refuses a stale marker (5 < 6) with a rebuild hint", raised)
    accepted = True
    try:
        ctc.require_current_normalization(cur, "x.xlsx", 5, "at-version")
        ctc.require_current_normalization(cur, "x.xlsx", 4, "above-version")
    except ValueError:
        accepted = False
    check("require accepts an at/above-version marker", accepted)


def test_mirror_invariant():
    print("catalog normalization_version MIRRORS the comparator constant:")
    for key, mod, name in (("ramp_detail", rd, "Ramp Detail"),
                           ("intersection_detail", idt, "Intersection Detail"),
                           ("highway_detail", hdt, "Highway Detail")):
        cat = tsn_library.get(key).normalization_version
        check(f"{name}: catalog {cat} == comparator {mod.NORMALIZATION_VERSION}",
              cat == mod.NORMALIZATION_VERSION)


def _project(_snapshot_path):
    def make_result(out_name):
        return ConsolidateResult(status="ok", message="ok",
                                 summary_lines=[out_name],
                                 completion=outcome.COMPLETE,
                                 skipped_inputs=0, failed_inputs=0)
    return [["001", "a"]], make_result


def _build(raw, out, marker_version):
    return tsn_library.build_normalized(
        str(raw), str(out), glob="*.xlsx", deps_ok=True, deps_msg="deps",
        no_raw_what="X", no_raw_hint="hint", log_label="Test", sheet="Data",
        header=["Route", "A"], header_align={"horizontal": "center"},
        project=_project, marker_version=marker_version)


def test_build_normalized_stamps():
    print("build_normalized stamps the requested marker version:")
    with tempfile.TemporaryDirectory() as d:
        d = Path(d)
        raw = d / "raw"
        raw.mkdir()
        src = Workbook()
        src.active.append(["x"])
        src.save(raw / "src.xlsx")
        src.close()

        out = d / "norm.xlsx"
        res = _build(raw, out, 9)
        check("build_normalized returned ok", getattr(res, "status", None) == "ok")
        wb = load_workbook(out, read_only=True, data_only=True)
        try:
            check("the built workbook carries the marker == 9",
                  ctc.normalization_marker_version(wb) == 9)
            check("the data sheet survives beside the marker",
                  "Data" in wb.sheetnames
                  and ctc.NORMALIZATION_MARKER_SHEET in wb.sheetnames)
        finally:
            wb.close()

        out2 = d / "norm2.xlsx"
        _build(raw, out2, None)
        wb2 = load_workbook(out2, read_only=True, data_only=True)
        try:
            check("marker_version=None writes NO marker sheet (RS/IS callers)",
                  ctc.NORMALIZATION_MARKER_SHEET not in wb2.sheetnames)
        finally:
            wb2.close()


def test_shared_header_helper():
    print("require_shared_header_prefix helper (CMP-AUD-033):")
    prefix = ["Route", "County", "PM", "Desc"]
    sidecars = ("TSN District", "TSN County")
    ok_hdr = prefix + list(sidecars)
    ctc.require_shared_header_prefix(ok_hdr, prefix, sidecars, "x.xlsx", "R")
    check("exact prefix + documented sidecars accepts", True)
    ctc.require_shared_header_prefix(["Route", " County ", "PM", "Desc"] + list(sidecars),
                                     prefix, sidecars, "x.xlsx", "R")
    check("whitespace on the header cells is tolerated", True)

    def refused(hdr):
        try:
            ctc.require_shared_header_prefix(hdr, prefix, sidecars, "x.xlsx", "R")
            return False
        except ValueError as e:
            return "rebuild" in str(e)
    check("a reordered prefix refuses",
          refused(["Route", "PM", "County", "Desc"] + list(sidecars)))
    check("a renamed prefix column refuses",
          refused(["Route", "Cnty", "PM", "Desc"] + list(sidecars)))
    check("a missing sidecar refuses", refused(prefix + ["TSN District"]))
    check("an undocumented trailing column refuses", refused(ok_hdr + ["Bonus"]))
    check("a shared column duplicated among the sidecars refuses",
          refused(ok_hdr + ["PM"]))


def _reordered_norm(sheet, shared, sidecars, path):
    swapped = [shared[1], shared[0]] + list(shared[2:])   # swap the first two shared cols
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Route"] + swapped + list(sidecars))
    ws.append(["001"] + ["x"] * (len(swapped) + len(sidecars)))
    ctc.write_normalization_marker(wb, 99)                # pass any marker gate; isolate the reorder
    wb.save(path)
    wb.close()


def test_loaders_bind_header():
    print("each normalized loader binds its header before reading (CMP-AUD-033):")
    fams = [("RD", rd, rd.NORMALIZED_SHEET, rd._NORMALIZED_SIDECARS),
            ("ID", idt, idt.NORMALIZED_SHEET, idt._NORMALIZED_SIDECARS),
            ("HD", hdt, hdt.NORMALIZED_SHEET, hdt._NORMALIZED_SIDECARS),
            ("HSL", hsl, hsl.tsn_hsl.NORMALIZED_SHEET, ())]
    with tempfile.TemporaryDirectory() as d:
        for tag, mod, sheet, sidecars in fams:
            p = Path(d) / f"{tag}_reordered.xlsx"
            _reordered_norm(sheet, list(mod.SHARED_HEADER), sidecars, p)
            try:
                mod._load_tsn(str(p))
                check(f"{tag} _load_tsn refuses a reordered shared header", False)
            except ValueError as e:
                check(f"{tag} _load_tsn refuses a reordered shared header",
                      "column layout does not match" in str(e) and "rebuild" in str(e))


def test_sidecar_mirror():
    print("comparator _NORMALIZED_SIDECARS mirrors the loader SIDECAR_HEADER:")
    for tag, mod, loader in (("RD", rd, _load_rd), ("ID", idt, _load_id),
                             ("HD", hdt, _load_hd)):
        check(f"{tag}: comparator sidecars == loader SIDECAR_HEADER",
              list(mod._NORMALIZED_SIDECARS) == list(loader.SIDECAR_HEADER))


def main():
    test_helper()
    test_mirror_invariant()
    test_build_normalized_stamps()
    test_shared_header_helper()
    test_loaders_bind_header()
    test_sidecar_mirror()
    print()
    if _fail:
        print(f"{len(_fail)} CHECK(S) FAILED:")
        for f in _fail:
            print("  -", f)
        sys.exit(1)
    print("ALL TSN-NORMALIZATION-MARKER CHECKS PASSED")


if __name__ == "__main__":
    main()
