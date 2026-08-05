"""RB4-A1 — the one executable acceptance run for hotfix bundle RB-4 (HF-05 + HF-10).

Drives the exact end-user paths the GUI workers drive — matrix.build_comparison
(Everything vs-TSN / SELF / ENV cells with the worker's own owned-dir leases and
commit guard), day_matrix.build_day_cell (By Day vs-TSN), the on-demand cameras
(matrix.evidence_for_cell / day_matrix.evidence_for_day_cell), and the classic
Compare tab + PDF-vs-Excel by-day silent controls — against the frozen 2026-07-23
and 2026-07-09 pulls, and retains one machine-readable, hash-bound result set.

Phases (run in order; `--phase all` for base-or-head sweeps is deliberate NOT
offered — each phase is invoked explicitly so the retained logs name what ran):

  provision   copy the frozen inputs into per-side acceptance stores (base/head)
              and the TSN library copy, hashing every provisioned file
  generate    run the full cell set for one side (--side base|head) against one
              scripts tree (--tree; defaults to this repo). The base side runs
              against a `git archive` export of the recorded base commit so the
              pre-fix defect signatures bind to that exact runtime
  census      app-free recount of the retained audit evidence sets and the
              base-side sets: truncation population, read-set composition,
              prose declarations (the PCOA-FINAL-004/-006 defect signatures)

Every result lands under the acceptance root as JSON; the final manifest/verifier
(rb4-verify-manifest.py at the audit docs root) binds the whole set to the
acceptance head. This driver is acceptance tooling, not app runtime: it prints.
"""
import argparse
import hashlib
import json
import os
import shutil
import sys
import time
import traceback
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]

# --------------------------------------------------------------------------- #
# fixed identities
# --------------------------------------------------------------------------- #
RUN_ID_DEFAULT = "RB4-A1"
ROOT_DEFAULT = Path(r"C:\Users\Yunus\Downloads\TSMIS\_scratch\post-comparison-hotfixes\HF-05\rb4-a1")
HF10_RETAIN = Path(r"C:\Users\Yunus\Downloads\TSMIS\_scratch\post-comparison-hotfixes\HF-10\rb4-a1")

FROZEN = {
    "pull_2026_07_23": Path(r"C:\Users\Yunus\Downloads\TSMIS\_scratch\post-comparison-output-audit-claude-independent-2026-07-23\raw-extract\2026-07-23 ssor-prod"),
    "pull_2026_07_09_ssor": Path(r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9\2026-07-09 ssor-prod"),
    "pull_2026_07_09_ars": Path(r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9\2026-07-09 ars-prod"),
    "tsn_library": Path(r"C:\Users\Yunus\Downloads\TSMIS\tsn_library"),
}
AUDIT_EVERYTHING = Path(r"C:\Users\Yunus\Downloads\TSMIS\_scratch\post-comparison-output-audit-claude-independent-2026-07-23\everything-dest\comparisons\tsn")
AUDIT_BYDAY = Path(r"C:\Users\Yunus\Projects\TSMIS-Reports-Exporter\output\comparisons\tsn-by-day\2026-07-23 ssor-prod")

# The everything-store provisioning — the Stage 1B layout the audit measured:
# ssor-prod <- the 2026-07-23 pull, ssor-test <- the 2026-07-09 ssor pull,
# ars-prod <- the 2026-07-09 ars outage-substitute (ID/IS only).
# highway_detail is EXCLUDED everywhere: pre-release (owner 2026-07-21), its
# artifacts are not ground truth and RB-4 must not generate from them.
ENV_PROVISION = {
    "ssor-prod": ("pull_2026_07_23", (
        "highway_log", "highway_log_pdf", "highway_sequence",
        "highway_sequence_pdf", "intersection_detail", "intersection_detail_pdf",
        "intersection_summary", "intersection_summary_pdf",
        "ramp_detail", "ramp_detail_pdf", "ramp_summary", "ramp_summary_excel")),
    "ssor-test": ("pull_2026_07_09_ssor", (
        "highway_log", "highway_log_pdf", "highway_sequence",
        "highway_sequence_pdf", "intersection_detail", "intersection_detail_pdf",
        "intersection_summary", "ramp_detail", "ramp_detail_pdf",
        "ramp_summary", "ramp_summary_excel")),
    "ars-prod": ("pull_2026_07_09_ars", (
        "intersection_detail", "intersection_detail_pdf",
        "intersection_summary", "intersection_summary_pdf")),
}
# By Day inputs: one run folder (the frozen 2026-07-23 pull), same subdir set.
BYDAY_DAY, BYDAY_SOURCE = "2026-07-23", "ssor-prod"
TSN_LIB_REPORTS = ("highway_log", "highway_sequence", "intersection_detail",
                   "intersection_summary", "ramp_detail", "ramp_summary")

# The complete evidence-relevant cell set (highway_detail excluded, pre-release).
TSN_ROWS = ("highway_log", "highway_log_pdf", "highway_sequence",
            "highway_sequence_pdf", "intersection_detail",
            "intersection_detail_pdf", "ramp_detail", "ramp_detail_pdf")
# The self mode rides each PDF row (`vs_excel`); Highway Log's Excel row also
# offers the symmetric `vs_pdf` — exactly the audit's observed self sets.
SELF_CELLS = (("highway_log", "vs_pdf"),) + tuple(
    (r, "vs_excel") for r in TSN_ROWS if r.endswith("_pdf"))
# The five Everything ENV PDF-vs-PDF cells (PCOA-FINAL-007), paired the way the
# audit measured them: ID-PDF against the ars-prod outage substitute, the rest
# against the 2026-07-09 ssor pull provisioned as ssor-test.
ENV_CELLS = (("highway_log_pdf", "ssor-test"),
             ("highway_sequence_pdf", "ssor-test"),
             ("ramp_detail_pdf", "ssor-test"),
             ("ramp_summary", "ssor-test"),
             ("intersection_detail_pdf", "ars-prod"))
BASELINE = "ssor-prod"
EVIDENCE_REQUEST = {"enabled": True, "examples": 2, "layout": "both"}

# --------------------------------------------------------------------------- #
# the DECLARED populations
#
# Every asserting phase compares what it found against these constants, never
# against whatever happens to be on disk. A silently-missing evidence set is
# precisely the defect this driver exists to catch, and a phase that globs its
# own population can never see one: the retained RB4-A1 validate record found
# THREE sets, reported zero problems, and passed.
#
# DERIVED, not guessed — and cross-checked three independent ways:
#   * `visual_evidence.rows()` names the ten evidence-capable rows (five report
#     families x two editions). Highway Detail is excluded everywhere here
#     (pre-release, owner 2026-07-21), leaving 4 families x 2 editions = the 8
#     TSN_ROWS.
#   * `report_catalog.MATRIX`'s `self_id` column declares six self placements
#     (highway_log/vs_pdf plus five *_pdf/vs_excel); minus highway_detail_pdf
#     that is exactly the 5 SELF_CELLS. This is the PDF-vs-Excel comparison
#     MODE inside the Everything matrix — NOT the by-day PvE matrix, which is a
#     silent control: `pdf_excel_matrix` has its own five rows and must publish
#     no evidence at all (asserted separately as `pve_stray`).
#   * By Day runs the same 8 rows against the one frozen day.
#   8 + 5 + 8 = 21, and the base side's own on-disk manifests count 21 exactly
#   (13 under store/comparisons/tsn + 8 under the by-day folder).
# The cross-environment population is `visual_evidence.env_rows()` == the five
# PCOA-FINAL-007 placements in ENV_CELLS. So the derivation AGREES with the
# declared 21 + 5; nothing here is bent to reach those numbers.
#
# The env sets are HEAD-ONLY by design: the HF-10 lane is what the hotfix adds,
# and the base side's job is to prove those five sets do NOT exist there. So at
# base they are declared FORBIDDEN, not merely absent.
# --------------------------------------------------------------------------- #
EXPECTED_TSN_SETS = 8            # len(TSN_ROWS)
EXPECTED_SELF_SETS = 5           # len(SELF_CELLS)
EXPECTED_BYDAY_SETS = 8          # len(TSN_ROWS), for the one frozen day
EXPECTED_EVIDENCE_SETS = 21      # the three above — required on BOTH sides
EXPECTED_ENV_SETS = 5            # len(ENV_CELLS) — required at head, forbidden at base
# One camera per vs-TSN row in each of the two lanes, plus the env cameras at
# head only (phase_cameras runs the env loop under `if side == "head"`).
EXPECTED_CAMERA_RUNS_BASE = 16   # 2 * len(TSN_ROWS)
EXPECTED_CAMERA_RUNS_HEAD = 21   # + len(ENV_CELLS)
RESULT_OK_STATUS = "ok"          # ConsolidateResult's only success terminal
# The panel-text bounds, split because they mean opposite things.
# PRE-fix the Excel panel drew value[:26] with NO ellipsis, so any value past
# that bound was drawn WRONG — the PCOA-FINAL-004 defect population.
PREFIX_PANEL_TRUNCATION_LIMIT = 26
# POST-fix the panel draws a visibly elided prefix past `visual_evidence.
# PANEL_TEXT_MAX`. Legal elision is NOT a defect. Mirrored here because the
# census phase runs app-free (it also censuses AUDIT roots that predate this
# tree) and cannot import visual_evidence; phase_validate does have the module
# and asserts the two agree, so the mirror cannot drift silently.
PANEL_TEXT_MAX_DECLARED = 120
# `evidence_manifest.STATE_RENDERED` — the one state that publishes a workbook.
# A no_differences / no_examples generation writes a manifest and NO workbook
# (verified against the retained base side: 21 manifests, 17 workbooks).
EVIDENCE_STATE_RENDERED = "rendered"
MIN_EMBEDDED_PICTURES = 1        # a rendered evidence workbook embeds images
LEDGER_SHEET = "Ledger"
SUMMARY_SHEET = "Summary"
MAX_PROBLEMS_PRINTED = 20
# Excel's own repair log basename shape, written beside a repaired workbook.
EXCEL_REPAIR_LOG_GLOB = "error*.xml"
# Stated in the retained record so no reader can mistake what the Excel phase
# proves. DisplayAlerts=False suppresses the repair prompt, Excel offers no COM
# property for "this file was repaired", and the repair-log sweep is best
# effort — so this run CANNOT claim zero repairs, only that Open() did not
# raise. Anyone who needs a real repair verdict has to open the books with
# alerts ENABLED and watch for the prompt.
EXCEL_REPAIR_NOTE = (
    "NOT DETECTED. DisplayAlerts=False means Excel repairs a damaged workbook "
    "silently and Open() still returns, and Excel exposes no COM property for "
    "'this file was repaired'. `opened`/`open_without_error` mean only that "
    "Open() did not raise; they are NOT proof of zero repairs. `repair_logs` "
    "is a best-effort sweep for Excel's own log file and an empty list proves "
    "nothing.")

# --------------------------------------------------------------------------- #
# the base-red contract (phase_checks_at_base)
#
# A non-zero exit is NOT proof that a check detects the defect. It proves only
# that SOMETHING went wrong, and the two most common somethings are (a) the
# check crashed for an unrelated reason and (b) the check could not even LOAD
# against the base runtime — it referenced a function the hotfix adds, so it
# died before reaching its own assertion. Neither demonstrates red.
#
# So every check the driver runs is declared here with what it must do at base
# and, when it must be RED, the substring that must appear in its captured
# output. `checks` is derived from this map, so the run list and the contract
# cannot drift apart.
#
# SIGNATURE PROVENANCE: measured from the retained RB4-A1 base-red record
# (results/base-red-checks.json, taken at head df31e1e). The bundle's checks
# have been edited since (d57820d, a510f11, ab827b6), so these signatures MUST
# be re-confirmed against the final acceptance head before the base-red phase
# is treated as evidence. A signature that has drifted classifies its check
# `inconclusive` and fails the phase — loudly, which is the point.
# --------------------------------------------------------------------------- #
CHECK_RED = "red"                # failed at base WITH its declared signature
CHECK_GREEN = "green"            # passed at base — a control, not a detector
CHECK_INCONCLUSIVE = "inconclusive"   # failed, but proves nothing about the defect
# One of these on STDERR means the run ended on a load-class error. That alone
# does NOT make a check inconclusive: a check that covers both a CORRECTED
# behaviour on an existing API and a NEW API will legitimately print its
# assertion failures and only then die reaching for the function the hotfix
# adds. It is inconclusive only when it printed no signature at all — i.e. it
# never got far enough to assert anything.
LOAD_ERROR_MARKERS = ("ImportError:", "ModuleNotFoundError:", "AttributeError:")
# Every check RB-4 modified is declared here: what it must do at the base tree,
# the STDOUT substring that must appear when it must be red, and why.
#
# MEASURED, NOT GUESSED. Read off actual `--phase checks-at-base` runs against
# the bound base tree (2026-08-05, head cd46cd2), and reproduced on a re-run.
# Re-run that phase and re-read its output whenever a check file changes — a
# signature bound to a stale run is worse than no signature at all, because it
# still reads as evidence.
#
# The split it records: EIGHT of the ten checks RB-4 touched demonstrate the
# defect at base by printing their own named assertion failure, and two pass
# there as controls. An earlier reading of this same phase found only three
# reds, because five checks called an API the hotfix adds and died before
# asserting anything; cd46cd2 gave those five a stated-contract block so they
# now assert first and stop second. Their signatures below are those stated
# contracts, and several of them die on a load-class error AFTER printing —
# recorded as `died_after_signature`, never as a downgrade.
#
# `inconclusive` currently has NO members. It stays a declarable outcome, and
# its guard stays live: a check declared to assert nothing at base that starts
# printing FAIL lines has become red-capable, and the phase fails until it is
# re-declared red with a signature (`stale_inconclusive_problem`). A rule with
# no current members is exactly the kind that rots, so it is a named function
# with its own unit exercise rather than an inline condition.
BASE_CHECK_EXPECTATIONS = {
    # --- RED: these print their own assertion failure against the base
    # runtime. Two of them also die afterwards on an API the hotfix adds,
    # which is exactly what a check covering BOTH a corrected behaviour on an
    # existing API and a new function does. That is recorded as
    # `died_after_signature`, not treated as a downgrade: the defect was
    # demonstrated before the run ended.
    "check_evidence_excel_columns.py": {
        "expect": CHECK_RED,
        "signature": "FAIL: July District also resolves to its Location cell",
        "why": "at base the July-2026 edition resolves its fields off the wrong "
               "position map; the run then dies on the added _workbook_side",
    },
    "check_matrix_ownership.py": {
        "expect": CHECK_RED,
        "signature": "[FAIL] the frozen MODE is routed into evidence_for_cell",
        "why": "at base the env camera job does not carry its frozen mode_id "
               "through to evidence_for_cell",
    },
    "check_pdf_route_identity.py": {
        "expect": CHECK_RED,
        "signature": "FAIL visual_evidence._locate_env_sides exists",
        "why": "the base has no _locate_env_sides; this check asserts its "
               "presence rather than crashing on it, so it reports a real FAIL",
    },
    # --- RED via the cd46cd2 stated-contract blocks. Each names the contract
    # it depends on and asserts it FIRST, so the base runtime's lack of that
    # contract is reported as a named failure instead of a crash. Three of
    # these five still die afterwards (two on AttributeError, one on a
    # TypeError); `check_evidence_literal_cells` stops on a clean SystemExit
    # with no traceback at all, which needs no special handling — stdout is
    # what decides, and stderr simply carries nothing to classify.
    "check_visual_evidence.py": {
        "expect": CHECK_RED,
        "signature": "[FAIL] the read set is captured in labelled per-side "
                     "buckets, so a manifest names the document each side was "
                     "compared from",
        "why": "the base has no per-side read-set buckets, so a manifest "
               "cannot name the document each side was compared from",
    },
    "check_evidence_source_role.py": {
        "expect": CHECK_RED,
        "signature": "[FAIL] the engine addresses the compared workbook's rows "
                     "and renders each side's panel from it (_workbook_rows_at "
                     "+ _workbook_side)",
        "why": "the base cannot address the compared workbook's own rows, so a "
               "side is rendered from a borrowed print instead",
    },
    "check_evidence_manifest.py": {
        "expect": CHECK_RED,
        "signature": "[FAIL] the engine renders a side from the compared "
                     "WORKBOOK (_workbook_rows_at + _workbook_side), not from "
                     "a borrowed print",
        "why": "same missing workbook-addressing contract, seen from the "
               "manifest's side of the boundary",
    },
    "check_matrix.py": {
        "expect": CHECK_RED,
        "signature": "[FAIL] a cell comparison can be asked for evidence",
        "why": "the base build_cell_comparison takes no evidence request, so "
               "the matrix cannot decorate a cell it just built",
    },
    "check_evidence_literal_cells.py": {
        "expect": CHECK_RED,
        "signature": "[FAIL] a drawn panel string is full or visibly elided "
                     "(panel_cell_text)",
        "why": "the base panel draws value[:26] with no ellipsis, so a long "
               "cell is silently cut rather than visibly elided",
    },
    # --- GREEN controls: already-correct paths whose new assertions hold at
    # base too. A control that goes red at base means the declaration is
    # wrong, and that fails the phase as well.
    "check_pdf_excel_matrix.py": {
        "expect": CHECK_GREEN, "signature": None,
        "why": "the by-day PDF-vs-Excel lane is already silent at base",
    },
    "check_evidence_bundle.py": {
        "expect": CHECK_GREEN, "signature": None,
        "why": "the bundle's zip/credential-scan contract already holds at base",
    },
}
CHECK_FAIL_LINES_KEPT = 40       # FAIL lines retained per check
CHECK_STDOUT_TAIL_LINES = 8
CHECK_STDERR_TAIL_LINES = 6
# The bundle must carry SOME red-at-base evidence. A base-red phase in which
# nothing is red proves only that the checks ran; it is the degenerate
# all-green state the whole phase exists to rule out.
MIN_RED_AT_BASE_CHECKS = 1

# --------------------------------------------------------------------------- #
# the CLASSIC Compare-tab silent control
#
# The classic tab's endpoint (`gui_compare_api._begin_compare`) cannot be driven
# headlessly: it blocks on a native save dialog. But the endpoint is a claim →
# dialog → launch wrapper, and the thing the silence claim is ABOUT is the
# comparator it calls. So both lanes are driven exactly the way the endpoint's
# `build` lambda calls them — `mod.compare(...)` for the file lane and
# `adapter.compare_folders(...)` for the folder lane, resolved through the same
# COMPARE_REPORTS registry the endpoint resolves through, with the same
# `mode`/`confirm_overwrite` surface. That substitution is recorded per lane as
# `proxied_by` so no reader can take the record as "the endpoint ran".
# --------------------------------------------------------------------------- #
CLASSIC_PROXY_NOTE = (
    "the classic Compare tab endpoint blocks on a native save dialog and "
    "cannot run headlessly; the COMPARATOR it calls is driven directly, the "
    "same call gui_compare_api._begin_compare's build lambda makes")
CLASSIC_CONTROL_DIRNAME = "classic-control"
CLASSIC_CONTROL_MODE = "both"    # what the GUI sends with both outputs ticked
# One lane each, keyed by stable comparison-op key, paired with the Everything
# placement whose recorded provenance supplies the SAME two inputs the matrix
# used — so the control needs no path re-derivation of its own. Highway
# Sequence is chosen for both because it is the smallest corpus: this control
# is about SILENCE, not about count coverage.
CLASSIC_FILE_LANE = ("cmp:highway_sequence:tsn", "everything-tsn|highway_sequence")
CLASSIC_FOLDER_LANE = ("cmp:highway_sequence_pdf:env",
                       "everything-env|highway_sequence_pdf")
EVIDENCE_PROBE_GLOB = "*evidence*"
PLANTED_CONTROL_NAME = "planted (evidence).xlsx"
PLANTED_CONTROL_BYTES = b"planted positive control"

# The comparison workbook's count contract, mirrored from
# `artifact_store.comparison_counts` for the COM (live-formulas) reader. Both
# readers locate the two columns by UNIQUE EXACT LABEL, never by position.
COMPARISON_SHEET = "Comparison"
STATUS_LABEL = "Status"
DIFFS_LABEL = "Diffs"
BOTH_STATUS = "Both"
# Rows per COM round trip when reading a live-formulas twin. A whole-sheet
# `.Value` read is one round trip but materializes every cell at once, and a
# real comparison sheet is millions of cells (the measured Highway Log one is
# 63,710 rows x 41 columns = 2.6M).
COM_ROW_CHUNK = 2000

# --------------------------------------------------------------------------- #
# the BASE tree's runtime identity
#
# The base tree is a plain COPY of the base commit's runtime, not a git
# checkout, so `tree_stamp` finds no head and used to record `tree_commit:
# null`. That left every base-side result unbound — and with an unbound base,
# base-vs-head count invariance is trivially true if the two paths ever
# converge. The identity comes instead from the committed content binding,
# which compares every runtime file in base-tree/ against the base commit's own
# git blobs. The base commit SHA is read from that record, never spelled here.
# --------------------------------------------------------------------------- #
BASE_BINDING_FILE = "base-tree-binding.json"

_CHUNK = 1 << 20


def sha256_of(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(_CHUNK), b""):
            h.update(chunk)
    return h.hexdigest()


def _first_line(e):
    """`type(e).__name__` + the FIRST line of the message — the minimum every
    swallowed exception owes the log ("one log upload answers it")."""
    lines = str(e).splitlines()
    return f"{type(e).__name__}: {lines[0] if lines else ''}"


def _norm(path):
    """One spelling for a path used as a population key, so a declared path and
    a discovered one compare equal."""
    return str(Path(path).resolve())


def assert_declared_counts():
    """The declared populations must equal what the driver actually enqueues.

    A row added to TSN_ROWS / SELF_CELLS / ENV_CELLS without updating the
    constants would otherwise shrink or inflate an expectation silently — the
    same fail-open class the population check exists to remove. Raises rather
    than warning, and names every constant that disagrees."""
    declared = (
        (EXPECTED_TSN_SETS, len(TSN_ROWS), "EXPECTED_TSN_SETS"),
        (EXPECTED_SELF_SETS, len(SELF_CELLS), "EXPECTED_SELF_SETS"),
        (EXPECTED_BYDAY_SETS, len(TSN_ROWS), "EXPECTED_BYDAY_SETS"),
        (EXPECTED_ENV_SETS, len(ENV_CELLS), "EXPECTED_ENV_SETS"),
        (EXPECTED_EVIDENCE_SETS,
         EXPECTED_TSN_SETS + EXPECTED_SELF_SETS + EXPECTED_BYDAY_SETS,
         "EXPECTED_EVIDENCE_SETS"),
        (EXPECTED_CAMERA_RUNS_BASE, 2 * len(TSN_ROWS),
         "EXPECTED_CAMERA_RUNS_BASE"),
        (EXPECTED_CAMERA_RUNS_HEAD, 2 * len(TSN_ROWS) + len(ENV_CELLS),
         "EXPECTED_CAMERA_RUNS_HEAD"),
    )
    bad = [f"{name} is {want} but the cell set gives {got}"
           for want, got, name in declared if want != got]
    if bad:
        raise SystemExit("declared population disagrees with the cell set: "
                         + "; ".join(bad))


def population_diff(required, discovered, forbidden=None):
    """Compare a DECLARED population against what was actually found on disk.

    `required` and `forbidden` map identity -> the one path that identity owns;
    `discovered` is every path found. Returns
    {missing, extra, forbidden_present, duplicate}, each a sorted list of
    human-readable strings naming exactly which identities are involved.

    `duplicate` catches BOTH a path discovered more than once and a path
    claimed by two declared identities — a declaration bug in which one file
    would quietly satisfy two expectations. Nothing here touches the
    filesystem, so it is directly unit-exercisable."""
    forbidden = dict(forbidden or {})
    owner, duplicate = {}, []
    for ident, path in list(required.items()) + list(forbidden.items()):
        prior = owner.get(path)
        if prior is None:
            owner[path] = ident
        else:
            duplicate.append(f"{path} is claimed by both {prior} and {ident}")
    seen = {}
    for path in discovered:
        seen[path] = seen.get(path, 0) + 1
    duplicate += [f"{path} was discovered {n} times"
                  for path, n in seen.items() if n > 1]
    found = set(seen)
    return {
        "missing": sorted(f"{ident} ({path})"
                          for ident, path in required.items()
                          if path not in found),
        "extra": sorted(path for path in found if path not in owner),
        "forbidden_present": sorted(f"{ident} ({path})"
                                    for ident, path in forbidden.items()
                                    if path in found),
        "duplicate": sorted(duplicate),
    }


def classify_base_check(exit_code, stdout, stderr, signature):
    """The verdict for one base-tree check run, as a dict:
    {classification, signature_found, died_after_signature, load_error, reason}.

    green         exited 0 — a declared control, detecting nothing at base.
    red           the declared failure signature appears in the check's own
                  STDOUT. The defect WAS demonstrated, and that is what
                  red-at-base means — even if the run later died on a
                  load-class error. A check that covers both a corrected
                  behaviour on an existing API and a new API does exactly that:
                  it prints its FAIL lines, then dies reaching for the function
                  the hotfix adds. `died_after_signature` records it rather
                  than hiding it.
    inconclusive  the signature never appeared. Either the check could not
                  reach its own assertions at all (a load-class error on
                  stderr) or it failed for some other, undeclared reason.
                  Never counted as red.

    Signature matching is against STDOUT — the check's own reporting channel.
    A stack trace on stderr is how a run ENDED, not what it asserted.

    Pure: takes the captured text, touches nothing."""
    out, err = stdout or "", stderr or ""
    signature_found = bool(signature) and signature in out
    load_error = next((m.rstrip(":") for m in LOAD_ERROR_MARKERS if m in err),
                      None)
    verdict = {"signature_found": signature_found,
               "died_after_signature": bool(signature_found and load_error),
               "load_error": load_error}
    if exit_code == 0:
        return dict(verdict, classification=CHECK_GREEN, reason="exited 0")
    if signature_found:
        reason = "printed its declared signature"
        if load_error:
            reason += (f", then died on {load_error} — expected for a check "
                       "that also covers an API the hotfix adds")
        return dict(verdict, classification=CHECK_RED, reason=reason)
    if load_error:
        return dict(verdict, classification=CHECK_INCONCLUSIVE,
                    reason=(f"printed no signature and died on {load_error} — "
                            "it never reached its own assertions against the "
                            "base runtime"))
    return dict(verdict, classification=CHECK_INCONCLUSIVE,
                reason="failed without its declared signature, so it failed "
                       "for some other reason")


def stale_inconclusive_problem(name, expectation, fail_lines):
    """The declared-inconclusive guard, or None.

    A check declared to assert nothing at base HAS become red-at-base evidence
    the moment it prints a FAIL line, and leaving it parked under
    `inconclusive` would understate the bundle's own proof.

    Its own function, not an inline condition, because the inconclusive set is
    currently EMPTY — a rule with no members is exactly the kind that rots, so
    it stays directly exercisable."""
    if expectation != CHECK_INCONCLUSIVE or not fail_lines:
        return None
    return (f"{name}: declared inconclusive (it asserts nothing at base), but "
            f"it printed {len(fail_lines)} FAIL line(s) — re-declare it red "
            f"with a signature: {fail_lines[:2]}")


def write_json(path, payload):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True,
                               ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"  wrote {path}")


def tree_stamp(tree):
    """The tree-under-test's exact git head + runtime dirty state — the
    self-stamp every retained result carries so exact-head identity lives in
    the record itself (the RB-2 lesson), not only in the manifest's assertion.
    An exported base tree has no .git: (None, None). Dirtiness is judged over
    the RUNTIME roots only, so docs-side edits never poison a head stamp."""
    import subprocess
    try:
        head = subprocess.run(("git", "-C", str(tree), "rev-parse", "HEAD"),
                              capture_output=True, text=True)
        if head.returncode != 0:
            return None, None
        status = subprocess.run(
            ("git", "-C", str(tree), "status", "--porcelain", "--",
             "scripts", "version.py", "requirements.txt", "build"),
            capture_output=True, text=True)
        dirty = (bool(status.stdout.strip()) if status.returncode == 0
                 else None)
        return head.stdout.strip(), dirty
    except OSError:
        return None, None


def read_base_binding(root):
    """The committed binding that proves base-tree/ IS the base commit.

    Raises SystemExit when it is missing, unreadable, or does not bind: a base
    tree that cannot be tied to a commit cannot self-stamp, and a result that
    cannot say which runtime produced it is not evidence."""
    path = Path(root) / "results" / BASE_BINDING_FILE
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as e:
        raise SystemExit(
            f"the base-tree binding {path} could not be read "
            f"({_first_line(e)}); the base side cannot self-stamp its "
            "runtime, so no result from it is evidence") from e
    commit, tree = payload.get("base_commit"), payload.get("base_tree")
    if not payload.get("bound") or not commit or not tree:
        raise SystemExit(
            f"the base-tree binding {path} does NOT bind: "
            f"bound={payload.get('bound')!r} base_commit={commit!r} "
            f"base_tree={tree!r}")
    return {"base_commit": commit, "base_tree": tree, "bound": True,
            "files_compared": payload.get("files_compared"),
            "base_runtime_digest": payload.get("base_runtime_digest")}


def runtime_stamp(root, tree, side):
    """(tree_commit, tree_runtime_dirty, stamp_record) for the tree a phase is
    running against — the self-stamp every retained result carries.

    A git checkout stamps itself via `tree_stamp`. The BASE tree has no .git,
    so it is stamped from the content binding instead, and `not bound` is the
    base-side notion of dirty (a base tree whose content no longer matches the
    commit is exactly what a dirty flag is for) rather than a misleading False.

    Both directions are asserted: a base-side result may never carry the
    acceptance head's commit, and a head-side result may never carry the base
    commit or run out of the base tree. A violation raises — the whole point of
    the stamp is that a reader can trust which runtime produced the numbers."""
    binding = read_base_binding(root)
    tree_path = _norm(tree)
    is_base_tree = tree_path == _norm(binding["base_tree"])
    commit, dirty = tree_stamp(tree)
    record = {"side": side, "tree": tree_path,
              "tree_is_base_tree": is_base_tree, "git_commit": commit,
              "git_dirty": dirty, "base_binding": binding,
              "stamp_source": "content binding" if is_base_tree else "git"}
    if is_base_tree:
        commit, dirty = binding["base_commit"], not binding["bound"]
    problems = []
    if side == "base" and not is_base_tree:
        problems.append(f"side=base ran against {tree_path}, which is not the "
                        f"bound base tree {binding['base_tree']}")
    if side != "base" and is_base_tree:
        problems.append(f"side={side} ran against the BASE tree {tree_path}")
    if side != "base" and commit == binding["base_commit"]:
        problems.append(f"side={side} self-stamps the BASE commit {commit}")
    if not commit:
        problems.append(f"no runtime commit could be recorded for {tree_path}")
    if problems:
        raise SystemExit("runtime self-stamp refused: " + "; ".join(problems))
    record["tree_commit"], record["tree_runtime_dirty"] = commit, dirty
    return commit, dirty, record


class Tee:
    """Mirror stdout/stderr into the retained per-phase log."""

    def __init__(self, path):
        path.parent.mkdir(parents=True, exist_ok=True)
        self._f = open(path, "a", encoding="utf-8")
        self._stdout = sys.stdout

    def write(self, text):
        self._stdout.write(text)
        self._f.write(text)
        self._f.flush()

    def flush(self):
        self._stdout.flush()
        self._f.flush()


# --------------------------------------------------------------------------- #
# provision
# --------------------------------------------------------------------------- #
def _copy_tree_hashed(src, dst, inventory, source_key):
    dst.mkdir(parents=True, exist_ok=True)
    n = 0
    for p in sorted(src.rglob("*")):
        rel = p.relative_to(src)
        target = dst / rel
        if p.is_dir():
            target.mkdir(parents=True, exist_ok=True)
            continue
        shutil.copyfile(p, target)
        inventory.append({"path": str(target), "source": str(p),
                          "source_key": source_key,
                          "size": target.stat().st_size,
                          "sha256": sha256_of(target)})
        n += 1
    return n


def phase_provision(root, sides, tree):
    sys.path.insert(0, str(Path(tree) / "scripts"))
    import owned_dir
    t0 = time.time()
    inventory = []
    lib = root / "lib" / "tsn_library"
    if lib.exists():
        print(f"  lib already provisioned at {lib} — leaving as-is")
    else:
        for report in TSN_LIB_REPORTS:
            n = _copy_tree_hashed(FROZEN["tsn_library"] / report, lib / report,
                                  inventory, f"tsn_library/{report}")
            print(f"  lib {report}: {n} file(s)")
    for side in sides:
        store = root / side / "store"
        out = root / side / "output"
        for env, (src_key, subdirs) in ENV_PROVISION.items():
            # The Everything worker leases each cell's env store as an EXISTING
            # app-owned root (require_existing_owned_dir_lease), so the replica
            # store must be created exactly the way the app creates one.
            if owned_dir.ensure_owned_dir(store / env, kind="store") is None:
                raise SystemExit(f"could not create owned store dir {store / env}")
            for sub in subdirs:
                n = _copy_tree_hashed(FROZEN[src_key] / sub, store / env / sub,
                                      inventory, f"{src_key}/{sub}")
            print(f"  {side}/store/{env}: provisioned {len(subdirs)} subdir(s)")
        day_dir = out / f"{BYDAY_DAY} {BYDAY_SOURCE}"
        src_key, subdirs = ENV_PROVISION["ssor-prod"]
        for sub in subdirs:
            _copy_tree_hashed(FROZEN[src_key] / sub, day_dir / sub,
                              inventory, f"{src_key}/{sub}")
        print(f"  {side}/output/{day_dir.name}: provisioned {len(subdirs)} subdir(s)")
    write_json(root / "results" / "provision.json", {
        "run_id": RUN_ID_DEFAULT, "elapsed_s": round(time.time() - t0, 1),
        "sides": list(sides), "files": inventory, "count": len(inventory)})
    print(f"provision done: {len(inventory)} file(s) in {time.time() - t0:.0f}s")


# --------------------------------------------------------------------------- #
# generate — runs INSIDE the tree under test
# --------------------------------------------------------------------------- #
def _sandbox(tree, root, side):
    scripts = str(Path(tree) / "scripts")
    sys.path.insert(0, scripts)
    import paths, settings, day_matrix  # noqa: E401
    side_root = root / side
    paths.OUTPUT_ROOT = side_root / "output"
    day_matrix.OUTPUT_ROOT = side_root / "output"
    paths.TSN_LIBRARY_ROOT = root / "lib" / "tsn_library"
    cfg = root / "config" / f"config-{side}.json"
    cfg.parent.mkdir(parents=True, exist_ok=True)
    settings.CONFIG_FILE = cfg
    settings._cache = settings._cache_mtime = None
    (side_root / "output").mkdir(parents=True, exist_ok=True)
    return side_root


def _events(module_events, log_path):
    log_path.parent.mkdir(parents=True, exist_ok=True)
    f = open(log_path, "a", encoding="utf-8")

    def on_log(msg):
        line = f"[{time.strftime('%H:%M:%S')}] {msg}"
        print(line)
        f.write(line + "\n")
        f.flush()

    return module_events.Events(on_log=on_log, is_cancelled=lambda: False)


def _typed_counts(consolidation_meta, path):
    """The strict typed sidecar read the matrix itself trusts."""
    rec = consolidation_meta.read_comparison_outcome(path)
    if rec is None:
        return {"present": False}
    oc = rec.comparison_outcome
    out = {"present": True, "trusted": bool(rec.trusted),
           "current": bool(rec.current)}
    if oc is not None:
        for k in ("verdict", "completion", "pairing_quality"):
            out[k] = str(getattr(oc, k, None))
        counts = getattr(oc, "counts", None)
        if counts is not None:
            for k in ("known", "paired_rows", "side_a_only_rows",
                      "side_b_only_rows", "differing_rows", "differing_cells",
                      "asserted_cells", "context_cells"):
                out[k] = getattr(counts, k, None)
            out["per_field_counts"] = dict(getattr(counts, "per_field_counts",
                                                   {}) or {})
    gen = getattr(rec, "artifact_generation", None)
    out["generation_id"] = getattr(gen, "generation_id", None)
    return out


def _evidence_sibling_paths(comparison_path):
    """(workbook, image folder, manifest) beside one comparison — the app-free
    twin of `visual_evidence.sibling_paths`, so no phase re-spells the naming."""
    p = Path(comparison_path)
    return (p.with_name(f"{p.stem} (evidence){p.suffix}"),
            p.with_name(f"{p.stem} (evidence images)"),
            p.with_name(f"{p.stem} (evidence).json"))


def _placements(side_root, day_matrix):
    """Every comparison placement this acceptance drives, as one ordered
    enumeration: [{identity, kind, flavor, row, cell, mode, path}].

    The counts, validate and excel phases all derive their expected population
    from HERE, so the three can never disagree about what the run was supposed
    to produce. `identity` keeps the counts record's key idiom exactly (that
    shape is read downstream)."""
    store = side_root / "store"
    tsn_dir = store / "comparisons" / "tsn"
    out = []
    for row in TSN_ROWS:
        out.append({"identity": f"everything-tsn|{row}",
                    "kind": "everything-tsn", "flavor": "tsn", "row": row,
                    "cell": BASELINE, "mode": "tsn",
                    "path": tsn_dir / f"{BASELINE}_{row}_tsn.xlsx"})
    for row, mode_id in SELF_CELLS:
        out.append({"identity": f"everything-self|{row}|{mode_id}",
                    "kind": "everything-self", "flavor": "self", "row": row,
                    "cell": BASELINE, "mode": mode_id,
                    "path": tsn_dir / f"{BASELINE}_{row}_{mode_id}.xlsx"})
    for row, other_env in ENV_CELLS:
        out.append({"identity": f"everything-env|{row}",
                    "kind": "everything-env", "flavor": "env", "row": row,
                    "cell": other_env, "mode": "env",
                    "path": (store / "comparisons" / BASELINE
                             / f"{other_env}_{row}.xlsx")})
    for row in TSN_ROWS:
        out.append({"identity": f"byday|{row}", "kind": "byday",
                    "flavor": "tsn", "row": row,
                    "cell": f"{BYDAY_DAY} {BYDAY_SOURCE}", "mode": "tsn",
                    "path": Path(day_matrix.day_out_path(
                        BYDAY_DAY, BYDAY_SOURCE, row))})
    return out


def _probe_evidence(tree):
    """(probe_saw_planted, stray) for one destination tree.

    A bare "the glob found nothing" is not evidence of silence — it reads
    identically whether the lane is silent, the tree is the wrong one, or the
    pattern is broken. So the probe is CONTROLLED first: plant an
    evidence-named file where one would land, require the probe to SEE it,
    remove it, and only then read the lane's own silence. (Same pattern as
    build/check_pdf_excel_matrix.py.)"""
    planted = Path(tree) / PLANTED_CONTROL_NAME
    planted.write_bytes(PLANTED_CONTROL_BYTES)
    try:
        seen = sorted(p.name for p in Path(tree).rglob(EVIDENCE_PROBE_GLOB))
    finally:
        planted.unlink()
    saw_planted = seen == [PLANTED_CONTROL_NAME]
    stray = sorted(str(p) for p in Path(tree).rglob(EVIDENCE_PROBE_GLOB))
    return saw_planted, stray


def _classic_silent_control(side_root, ev, day_matrix):
    """Drive the CLASSIC Compare tab's own comparators and prove that lane is
    silent. Returns one record per lane (never raises for a lane failure — each
    lane's own `problems` list carries it, and phase_generate fails on them).

    See CLASSIC_PROXY_NOTE: the endpoint blocks on a native save dialog, so
    what runs is the comparator the endpoint calls, resolved through the same
    COMPARE_REPORTS registry. Each lane is fed the SAME two inputs the matrix
    fed its Everything placement, read back from that comparison's own
    provenance sidecar."""
    import compare_tsn_common as ctc
    from reports import COMPARE_REPORTS, compare_index_for_key

    places = {p["identity"]: p for p in _placements(side_root, day_matrix)}
    records = []
    for lane, (op_key, identity) in (("file", CLASSIC_FILE_LANE),
                                     ("folder", CLASSIC_FOLDER_LANE)):
        rec = {"lane": lane, "op_key": op_key, "source_placement": identity,
               "proxied_by": CLASSIC_PROXY_NOTE, "problems": []}
        records.append(rec)
        print(f"=== classic control · {lane} lane · {op_key} ===")
        index = compare_index_for_key(op_key)
        if index is None:
            rec["problems"].append(f"{op_key} is not a registered comparison")
            continue
        label, adapter, kind, _group = COMPARE_REPORTS[index][:4]
        rec["label"], rec["kind"] = label, kind
        place = places.get(identity)
        if place is None:
            rec["problems"].append(f"{identity} is not a known placement")
            continue
        prov = ctc.read_comparison_provenance(place["path"])
        if not isinstance(prov, dict):
            rec["problems"].append(
                f"no provenance sidecar for {identity} — this control reuses "
                "that placement's own two inputs, so its cell must be "
                "generated first")
            continue
        sides = [Path(s["selection"]) for s in prov["inputs"]]
        rec["inputs"] = [str(s) for s in sides]
        dest_dir = side_root / CLASSIC_CONTROL_DIRNAME / lane
        if dest_dir.exists():
            shutil.rmtree(dest_dir)
        dest_dir.mkdir(parents=True)
        out_path = dest_dir / f"classic_{lane}_control.xlsx"
        started = time.time()
        try:
            if kind == "files":
                res = adapter.compare(sides[0], sides[1], out_path, events=ev,
                                      confirm_overwrite=None,
                                      mode=CLASSIC_CONTROL_MODE)
            else:
                res = adapter.compare_folders(sides[0], sides[1], out_path,
                                              events=ev, confirm_overwrite=None,
                                              mode=CLASSIC_CONTROL_MODE)
            rec["status"] = res.status
            rec["message"] = getattr(res, "message", "")
        except Exception as e:  # noqa: BLE001 — recorded as this lane's problem
            rec["status"] = "exception"
            rec["error"] = f"{type(e).__name__}: {e}"
            rec["traceback"] = traceback.format_exc()
            print(rec["traceback"])
        rec["elapsed_s"] = round(time.time() - started, 1)
        if rec.get("status") != RESULT_OK_STATUS:
            reason = rec.get("error") or rec.get("message")
            rec["problems"].append(
                f"{lane} lane did not complete: status={rec.get('status')!r} "
                f"— {reason if reason else 'no reason recorded'}")
        rec["wrote"] = sorted(p.name for p in dest_dir.rglob("*")
                              if p.is_file())
        if not rec["wrote"]:
            rec["problems"].append(
                f"{lane} lane wrote nothing, so there was no tree to search "
                "and its silence proves nothing")
        saw_planted, stray = _probe_evidence(dest_dir)
        rec["probe_saw_planted"], rec["evidence_stray"] = saw_planted, stray
        if not saw_planted:
            rec["problems"].append(
                f"{lane} lane: the silence probe did NOT see its planted "
                "control, so its 'no evidence' reading is meaningless")
        if stray:
            rec["problems"].append(
                f"{lane} lane wrote evidence artifacts: {stray}")
        print(f"  classic {lane}: status={rec.get('status')} "
              f"files={len(rec['wrote'])} probe_ok={saw_planted} "
              f"stray={len(stray)}")
    return records


def _evidence_state(path):
    """App-free look at one comparison's evidence siblings."""
    wb, imgs, man = _evidence_sibling_paths(path)
    state = {"workbook": wb.exists(), "images": imgs.is_dir(),
             "manifest": man.exists(), "png_count": 0, "manifest_state": None,
             "read_set": None}
    if state["images"]:
        state["png_count"] = sum(1 for _ in imgs.glob("*.png"))
    if state["manifest"]:
        try:
            payload = json.loads(man.read_text(encoding="utf-8"))
            state["manifest_state"] = payload.get("state")
            state["read_set"] = [m.get("name") for m in payload.get("read_set", [])]
        except (OSError, ValueError) as e:
            state["manifest_state"] = f"unreadable: {type(e).__name__}"
    return state


def phase_generate(root, side, tree, run_id, kinds=None, label=""):
    side_root = _sandbox(tree, root, side)
    import events as events_mod
    import matrix
    import day_matrix
    import owned_dir
    import consolidation_meta
    import tsn_library

    dest = side_root / "store"
    tag = f"{side}-{label}" if label else side
    log_path = root / "logs" / f"generate-{tag}.log"
    ev = _events(events_mod, log_path)
    commit, dirty, stamp = runtime_stamp(root, tree, side)
    results = {"run_id": run_id, "side": side, "tree": str(tree), "label": label,
               "tree_commit": commit, "tree_runtime_dirty": dirty,
               "runtime_stamp": stamp,
               "python": sys.version.split()[0], "cells": [], "started": time.time()}

    def kind_on(kind):
        return kinds is None or kind in kinds

    # The Settings rebuild entry, once per dataset (the owner-plan's "TSN
    # libraries rebuild once"): a dataset with no consolidated workbook is
    # BUILT here — the matrix heals stale libraries but never creates one.
    for report in ("highway_log", "highway_sequence", "intersection_detail",
                   "ramp_detail"):
        try:
            res = tsn_library.build_consolidated(report, events=ev)
            print(f"  tsn build {report}: {res.status} — {res.message}")
        except Exception as e:  # noqa: BLE001 — recorded; cells then refuse
            print(f"  tsn build {report}: EXCEPTION {type(e).__name__}: {e}")

    comparisons_lease = owned_dir.require_owned_dir_lease(
        dest / matrix.COMPARISONS_DIRNAME, kind="comparisons")
    active = {"store": None}

    def target_guard(path=None, *, anchor_path=None, anchor_identity=None,
                     directory_identity=None):
        leases = [x for x in (comparisons_lease, active["store"]) if x is not None]
        if not leases or not all(x.is_current() for x in leases):
            return False
        if path is None:
            return True
        return any(x.is_safe_descendant(
            path, anchor_path=anchor_path, anchor_identity=anchor_identity,
            directory_identity=directory_identity) for x in leases)

    def run_cell(kind, row_key, cell_key, mode_id, fn):
        started = time.time()
        entry = {"kind": kind, "row": row_key, "cell": cell_key, "mode": mode_id}
        print(f"=== {side} · {kind} · {row_key} · {cell_key} · {mode_id} ===")
        try:
            res = fn()
            entry["status"] = res.status
            entry["message"] = getattr(res, "message", "")
            entry["summary_lines"] = list(getattr(res, "summary_lines", ()) or ())
        except Exception as e:  # noqa: BLE001 — recorded, run continues
            entry["status"] = "exception"
            entry["error"] = f"{type(e).__name__}: {e}"
            entry["traceback"] = traceback.format_exc()
            print(entry["traceback"])
        entry["elapsed_s"] = round(time.time() - started, 1)
        results["cells"].append(entry)
        write_json(root / "results" / f"generation-{tag}.json", results)
        return entry

    # 1) Everything vs-TSN cells (evidence toggle ON, worker-parity leases).
    for row in TSN_ROWS if kind_on("everything-tsn") else ():
        def build_tsn(row=row):
            active["store"] = owned_dir.require_existing_owned_dir_lease(
                dest / BASELINE, kind="store")
            try:
                return matrix.build_comparison(
                    str(dest), row, BASELINE, "tsn", BASELINE, events=ev,
                    tsn_files={}, also_formulas=True,
                    evidence=dict(EVIDENCE_REQUEST), commit_guard=target_guard)
            finally:
                active["store"] = None
        entry = run_cell("everything-tsn", row, BASELINE, "tsn", build_tsn)
        cmp_path = (dest / matrix.COMPARISONS_DIRNAME / "tsn"
                    / f"{BASELINE}_{row}_tsn.xlsx")
        entry["comparison_path"] = str(cmp_path)
        entry["typed"] = _typed_counts(consolidation_meta, cmp_path)
        entry["evidence"] = _evidence_state(cmp_path)
        write_json(root / "results" / f"generation-{tag}.json", results)

    # 2) Everything SELF cells (the toggle-driven self decoration).
    for row, mode_id in SELF_CELLS if kind_on("everything-self") else ():
        def build_self(row=row, mode_id=mode_id):
            active["store"] = owned_dir.require_existing_owned_dir_lease(
                dest / BASELINE, kind="store")
            try:
                return matrix.build_comparison(
                    str(dest), row, BASELINE, mode_id, BASELINE, events=ev,
                    tsn_files={}, also_formulas=True,
                    evidence=dict(EVIDENCE_REQUEST), commit_guard=target_guard)
            finally:
                active["store"] = None
        entry = run_cell("everything-self", row, BASELINE, mode_id, build_self)
        cmp_path = (dest / matrix.COMPARISONS_DIRNAME / "tsn"
                    / f"{BASELINE}_{row}_{mode_id}.xlsx")
        entry["comparison_path"] = str(cmp_path)
        entry["typed"] = _typed_counts(consolidation_meta, cmp_path)
        entry["evidence"] = _evidence_state(cmp_path)
        write_json(root / "results" / f"generation-{tag}.json", results)

    # 3) Everything ENV cells — the five PDF-vs-PDF placements (PCOA-FINAL-007).
    #    At base this proves the comparisons run and NO evidence artifact exists;
    #    at head it exercises the new lane end to end.
    for row, other_env in ENV_CELLS if kind_on("everything-env") else ():
        cmp_path = (dest / matrix.COMPARISONS_DIRNAME / BASELINE
                    / f"{other_env}_{row}.xlsx")
        if side == "head":
            # HF-10 criterion 3: the env comparison's counts are identical with
            # evidence on and off. The OFF pass runs FIRST so the final store
            # state (and every evidence binding) belongs to the ON pass.
            def build_env_off(row=row, other_env=other_env):
                return matrix.build_comparison(
                    str(dest), row, other_env, "env", BASELINE, events=ev,
                    tsn_files={}, also_formulas=True,
                    evidence=None, commit_guard=target_guard)
            entry = run_cell("everything-env-off", row, other_env, "env",
                             build_env_off)
            entry["comparison_path"] = str(cmp_path)
            entry["typed"] = _typed_counts(consolidation_meta, cmp_path)
            entry["evidence"] = _evidence_state(cmp_path)
            write_json(root / "results" / f"generation-{tag}.json", results)

        def build_env(row=row, other_env=other_env):
            return matrix.build_comparison(
                str(dest), row, other_env, "env", BASELINE, events=ev,
                tsn_files={}, also_formulas=True,
                evidence=dict(EVIDENCE_REQUEST), commit_guard=target_guard)
        entry = run_cell("everything-env", row, other_env, "env", build_env)
        entry["comparison_path"] = str(cmp_path)
        entry["typed"] = _typed_counts(consolidation_meta, cmp_path)
        entry["evidence"] = _evidence_state(cmp_path)
        write_json(root / "results" / f"generation-{tag}.json", results)

    # 4) By Day vs-TSN cells (no commit guard — the day lane is app-private).
    for row in TSN_ROWS if kind_on("byday") else ():
        def build_day(row=row):
            return day_matrix.build_day_cell(
                BYDAY_SOURCE, BYDAY_DAY, row, str(dest), ev, tsn_files={},
                also_formulas=True, evidence=dict(EVIDENCE_REQUEST))
        entry = run_cell("byday-tsn", row, f"{BYDAY_DAY} {BYDAY_SOURCE}", "tsn",
                         build_day)
        cmp_path = day_matrix.day_out_path(BYDAY_DAY, BYDAY_SOURCE, row)
        entry["comparison_path"] = str(cmp_path)
        entry["typed"] = _typed_counts(consolidation_meta, cmp_path)
        entry["evidence"] = _evidence_state(cmp_path)
        write_json(root / "results" / f"generation-{tag}.json", results)

    # 5) Silent control A — the by-day PDF-vs-Excel matrix (its own lane, its
    #    own tree). NOT a stand-in for the classic Compare tab: that lane is
    #    driven separately in 5b, because a control that was never run is not a
    #    control.
    import pdf_excel_matrix
    for fam_row in (("highway_log_pdf", "highway_sequence_pdf",
                     "intersection_detail_pdf", "ramp_detail_pdf")
                    if kind_on("pve") else ()):
        def build_pve(fam_row=fam_row):
            return pdf_excel_matrix.build_pve_cell(
                BYDAY_SOURCE, BYDAY_DAY, fam_row, str(dest), ev,
                also_formulas=False)
        entry = run_cell("pve-byday", fam_row, f"{BYDAY_DAY} {BYDAY_SOURCE}",
                         "vs_excel", build_pve)
        try:
            cmp_path = pdf_excel_matrix.day_out_path(BYDAY_DAY, BYDAY_SOURCE,
                                                     fam_row)
            entry["comparison_path"] = str(cmp_path)
            entry["evidence"] = _evidence_state(cmp_path)
        except Exception as e:  # noqa: BLE001
            entry["path_error"] = f"{type(e).__name__}: {e}"
        write_json(root / "results" / f"generation-{tag}.json", results)

    # 5b) Silent control B — the CLASSIC Compare tab's own two comparators,
    #     driven directly and probe-controlled. Recorded under its own key so
    #     the record never conflates it with the PvE control above.
    if kind_on("classic"):
        results["classic_control"] = _classic_silent_control(
            side_root, ev, day_matrix)
        write_json(root / "results" / f"generation-{tag}.json", results)

    # 6) Whole-tree evidence sweep: every evidence artifact under the side root,
    #    so absences are proved by enumeration, not assumption.
    hits = sorted(str(p) for p in side_root.rglob("*evidence*"))
    results["evidence_sweep"] = hits
    # The silent-lane controls ASSERT. Per-cell generation failures stay
    # recorded-and-continue as they always have (a partial re-run is a normal
    # operation), but a control that did not prove silence is a hard failure —
    # it is the only thing in this phase whose whole purpose is to be believed.
    control_problems = [p for rec in results.get("classic_control", ())
                        for p in rec["problems"]]
    results["control_problems"] = control_problems
    results["finished"] = time.time()
    results["elapsed_s"] = round(results["finished"] - results["started"], 1)
    write_json(root / "results" / f"generation-{tag}.json", results)
    print(f"generate {side} done in {results['elapsed_s']:.0f}s — "
          f"{len(results['cells'])} cell(s), {len(hits)} evidence path(s), "
          f"{len(control_problems)} control problem(s)")
    if control_problems:
        for p in control_problems[:MAX_PROBLEMS_PRINTED]:
            print("  PROBLEM:", p)
        raise SystemExit(1)


# --------------------------------------------------------------------------- #
# census — app-free defect-signature recount (PCOA-FINAL-004/-006)
# --------------------------------------------------------------------------- #
def _workbook_summary_rows(path):
    """(header_line3, legend_texts, [(field, route_key, va, vb)]) from one
    evidence workbook, via openpyxl only."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Summary"]
        rows = []
        line3 = ""
        for r, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if r == 3:
                line3 = str(row[0] or "")
            if r >= 6 and row and row[0]:
                field, key = str(row[0]), str(row[1] or "")
                if key.startswith("no verifiable example"):
                    continue
                rows.append((field, key, str(row[2] or ""), str(row[3] or "")))
        legends = []
        for name in wb.sheetnames:
            if name in ("Summary", "Ledger"):
                continue
            ws2 = wb[name]
            for r, row in enumerate(ws2.iter_rows(min_row=2, max_row=2,
                                                  values_only=True), 2):
                if row and row[0]:
                    legends.append(str(row[0]))
                break
        return line3, legends, rows
    finally:
        wb.close()


def _census_one_root(root_dir, label):
    sets = []
    for man in sorted(Path(root_dir).glob("*(evidence).json")):
        stem = man.name[: -len(" (evidence).json")]
        payload = json.loads(man.read_text(encoding="utf-8"))
        read_set = [m.get("name", "") for m in payload.get("read_set", [])]
        entry = {"root": label, "set": stem, "state": payload.get("state"),
                 "read_set_pdf": sum(1 for n in read_set
                                     if n.lower().endswith(".pdf")),
                 "read_set_xlsx": sum(1 for n in read_set
                                      if n.lower().endswith(".xlsx")),
                 "read_set": read_set, "examples": [],
                 "truncated_examples": [], "legally_elided_examples": [],
                 "blank_side_examples": [], "summary_line3": None,
                 "legends": []}
        wb = man.with_name(f"{stem} (evidence).xlsx")
        if wb.exists():
            line3, legends, rows = _workbook_summary_rows(wb)
            entry["summary_line3"] = line3
            entry["legends"] = sorted(set(legends))
            for field, key, va, vb in rows:
                item = {"field": field, "at": key, "va_len": len(va),
                        "vb_len": len(vb)}
                entry["examples"].append(item)
                # The PRE-FIX Excel panel draws value[:26] with no ellipsis, so
                # any value past that bound was drawn WRONG.
                if (len(va) > PREFIX_PANEL_TRUNCATION_LIMIT
                        or len(vb) > PREFIX_PANEL_TRUNCATION_LIMIT):
                    entry["truncated_examples"].append(
                        {"field": field, "at": key, "va": va, "vb": vb})
                # A STRICT SUBSET of the above (the elision bound is far higher
                # than the pre-fix cut): past it the post-fix panel draws a
                # visibly elided prefix, which is LEGAL and not a defect.
                if (len(va) > PANEL_TEXT_MAX_DECLARED
                        or len(vb) > PANEL_TEXT_MAX_DECLARED):
                    entry["legally_elided_examples"].append(
                        {"field": field, "at": key})
                # PCOA-FINAL-005: the compared value is blank on exactly ONE
                # side — the hardest panels to draw, since the blank side has
                # no glyph to box. Counted as a population with identities
                # rather than left to a spot check.
                if bool(va) != bool(vb):
                    entry["blank_side_examples"].append(
                        {"field": field, "at": key,
                         "blank_side": "a" if not va else "b"})
        sets.append(entry)
    return sets


def phase_cameras(root, side, tree, run_id):
    """The ON-DEMAND per-cell cameras — the other end-user evidence path: the
    Everything vs-TSN camera, the By Day camera, and (at head) the HF-10 env
    camera, each regenerating evidence for an EXISTING comparison under the
    freshness gates.

    ASSERTS its outcomes: the declared number of cameras must have run, and
    every one of them must have reached the success terminal. Recording an
    `exception` (or a `cancelled` / `error` ConsolidateResult) and moving on
    would let a camera phase in which every single camera threw still pass."""
    side_root = _sandbox(tree, root, side)
    import events as events_mod
    import matrix
    import day_matrix
    import owned_dir
    dest = side_root / "store"
    ev = _events(events_mod, root / "logs" / f"cameras-{side}.log")
    commit, dirty, stamp = runtime_stamp(root, tree, side)
    expected_runs = (EXPECTED_CAMERA_RUNS_HEAD if side == "head"
                     else EXPECTED_CAMERA_RUNS_BASE)
    results = {"run_id": run_id, "side": side, "tree": str(tree), "cells": [],
               "tree_commit": commit, "tree_runtime_dirty": dirty,
               "runtime_stamp": stamp,
               "expected_runs": expected_runs, "started": time.time()}
    lease = owned_dir.require_existing_owned_dir_lease(
        dest / matrix.COMPARISONS_DIRNAME, kind="comparisons")

    def guard(path=None, *, anchor_path=None, anchor_identity=None,
              directory_identity=None):
        if not lease.is_current():
            return False
        if path is None:
            return True
        return lease.is_safe_descendant(
            path, anchor_path=anchor_path, anchor_identity=anchor_identity,
            directory_identity=directory_identity)

    def run(kind, row, cell, fn):
        entry = {"kind": kind, "row": row, "cell": cell}
        print(f"=== camera {side} · {kind} · {row} · {cell} ===")
        started = time.time()
        try:
            res = fn()
            entry["status"] = res.status
            entry["message"] = getattr(res, "message", "")
        except Exception as e:  # noqa: BLE001
            entry["status"] = "exception"
            entry["error"] = f"{type(e).__name__}: {e}"
            entry["traceback"] = traceback.format_exc()
            print(entry["traceback"])
        entry["elapsed_s"] = round(time.time() - started, 1)
        results["cells"].append(entry)
        write_json(root / "results" / f"cameras-{side}.json", results)

    for row in TSN_ROWS:
        run("camera-tsn", row, BASELINE,
            lambda row=row: matrix.evidence_for_cell(
                str(dest), row, BASELINE, BASELINE, ev, tsn_files={},
                commit_guard=guard))
        run("camera-byday", row, f"{BYDAY_DAY} {BYDAY_SOURCE}",
            lambda row=row: day_matrix.evidence_for_day_cell(
                BYDAY_SOURCE, BYDAY_DAY, row, str(dest), ev, tsn_files={}))
    if side == "head":
        for row, other_env in ENV_CELLS:
            run("camera-env", row, other_env,
                lambda row=row, other_env=other_env: matrix.evidence_for_cell(
                    str(dest), row, other_env, BASELINE, ev, tsn_files={},
                    commit_guard=guard, mode_id="env"))
    problems = []
    if len(results["cells"]) != expected_runs:
        problems.append(
            f"{len(results['cells'])} camera run(s) recorded, but side "
            f"{side} declares {expected_runs}")
    for entry in results["cells"]:
        status = entry.get("status")
        if status == RESULT_OK_STATUS:
            continue
        # Never `or`-mask the reason: an outcome with neither an error nor a
        # message says so, rather than reporting an empty string.
        reason = entry.get("error") or entry.get("message")
        problems.append(
            f"camera {entry['kind']} · {entry['row']} · {entry['cell']}: "
            f"status={status!r} (expected {RESULT_OK_STATUS!r}) — "
            f"{reason if reason else 'no reason recorded'}")
    results["problems"] = problems
    results["problem_count"] = len(problems)
    results["elapsed_s"] = round(time.time() - results["started"], 1)
    write_json(root / "results" / f"cameras-{side}.json", results)
    print(f"cameras {side}: {len(results['cells'])}/{expected_runs} run(s), "
          f"{len(problems)} problem(s)")
    if problems:
        for p in problems[:MAX_PROBLEMS_PRINTED]:
            print("  PROBLEM:", p)
        raise SystemExit(1)


def _summary_rows_of(wb_path):
    """App-free read of one evidence workbook: (source_lines, legends,
    example rows [(field, at, va, vb)], has_ledger)."""
    from openpyxl import load_workbook
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    try:
        ws = wb["Summary"]
        src_lines, rows = [], []
        for r, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if r in (3, 4) and row and row[0]:
                src_lines.append(str(row[0]))
            if r >= 7 and row and row[0]:
                key = str(row[1] or "")
                if key.startswith("no verifiable example"):
                    continue
                rows.append((str(row[0]), key, str(row[2] or ""),
                             str(row[3] or "")))
        legends = []
        for name in wb.sheetnames:
            if name in ("Summary", "Ledger"):
                continue
            for row in wb[name].iter_rows(min_row=2, max_row=2,
                                          values_only=True):
                if row and row[0]:
                    legends.append(str(row[0]))
                break
        return src_lines, sorted(set(legends)), rows, "Ledger" in wb.sheetnames
    finally:
        wb.close()


def phase_counts(root, side, tree, run_id):
    """Re-read every placement's strict typed sidecar into one counts table —
    the base/head invariance record and the audit-count cross-check."""
    side_root = _sandbox(tree, root, side)
    import consolidation_meta
    import day_matrix
    commit, dirty, stamp = runtime_stamp(root, tree, side)
    out = {"run_id": run_id, "side": side, "tree": str(tree),
           "tree_commit": commit, "tree_runtime_dirty": dirty,
           "runtime_stamp": stamp, "cells": {}}
    for place in _placements(side_root, day_matrix):
        out["cells"][place["identity"]] = _typed_counts(
            consolidation_meta, place["path"])
    write_json(root / "results" / f"counts-{side}.json", out)


def phase_validate(root, side, tree, run_id):
    """The programmatic 100 % validation over every DECLARED evidence set:
    the expected population itself, then per set manifest member integrity,
    exact-source read-set binding against each comparison's own provenance,
    truthful source lines + legends, panel-text fidelity census, env geometry
    re-derivation, and the silent-lane sweep.

    The population check comes FIRST and is the reason this phase can be
    trusted at all. It used to glob whatever evidence sets happened to exist,
    which meant a run that produced three of twenty-six sets validated those
    three and reported zero problems — exactly what the retained RB4-A1 record
    shows. Every set is now driven from `_placements`, and a shortfall, an
    unexpected extra, a set that must NOT exist on this side, or a duplicated
    identity fails the phase naming the identities involved."""
    side_root = _sandbox(tree, root, side)
    import compare_env
    import compare_tsn_common as ctc
    import day_matrix
    import evidence_manifest as em
    import visual_evidence as ve

    problems = []
    commit, dirty, stamp = runtime_stamp(root, tree, side)
    out = {"run_id": run_id, "side": side, "tree": str(tree),
           "tree_commit": commit, "tree_runtime_dirty": dirty,
           "runtime_stamp": stamp, "sets": [], "problems": problems}

    # The app-free census mirrors the elision bound; if the app's own constant
    # moved, that census would mis-split legal elision from the pre-fix
    # truncation defect and quietly report the wrong population.
    if ve.PANEL_TEXT_MAX != PANEL_TEXT_MAX_DECLARED:
        problems.append(
            f"visual_evidence.PANEL_TEXT_MAX is {ve.PANEL_TEXT_MAX} but this "
            f"driver declares {PANEL_TEXT_MAX_DECLARED}; update "
            "PANEL_TEXT_MAX_DECLARED or the census totals are wrong")

    # The declared population. The 21 vs-TSN / self / by-day sets are required
    # on both sides; the 5 cross-environment sets are the HF-10 lane, so they
    # are required at head and FORBIDDEN at base (the base side's job is to
    # prove they do not exist there).
    places = _placements(side_root, day_matrix)
    required, forbidden = {}, {}
    for place in places:
        manifest = _evidence_sibling_paths(place["path"])[2]
        env_at_base = place["flavor"] == "env" and side != "head"
        (forbidden if env_at_base else required)[place["identity"]] = _norm(
            manifest)
    discovered = [_norm(p) for p in side_root.rglob("*(evidence).json")]
    pop = population_diff(required, discovered, forbidden)
    out["population"] = dict(
        pop, declared_evidence_sets=EXPECTED_EVIDENCE_SETS,
        declared_env_sets=EXPECTED_ENV_SETS, required=len(required),
        forbidden=len(forbidden), discovered=len(discovered))
    for label, items in (
            ("evidence set MISSING", pop["missing"]),
            ("UNEXPECTED evidence set (not a declared placement)", pop["extra"]),
            (f"evidence set that must NOT exist on side {side}",
             pop["forbidden_present"]),
            ("DUPLICATE population identity", pop["duplicate"])):
        problems.extend(f"{label}: {item}" for item in items)

    def note(setrec, cond, what):
        setrec.setdefault("checks", []).append(
            {"ok": bool(cond), "what": what})
        if not cond:
            problems.append(f"{setrec['comparison']}: {what}")

    def validate_set(cmp_path, flavor, row_key=None, identity=""):
        man_path = em.manifest_path(cmp_path)
        rec = {"comparison": str(cmp_path), "flavor": flavor,
               "identity": identity, "manifest": man_path.exists()}
        out["sets"].append(rec)
        if not man_path.exists():
            note(rec, False, "manifest missing")
            return
        man = em.read(man_path)
        rec["state"] = man.state
        desc = em.describe(cmp_path, verify_members=True)
        note(rec, desc["status"] == em.CURRENT,
             f"manifest describes CURRENT (got {desc['status']})")
        prov = ctc.read_comparison_provenance(cmp_path)
        note(rec, isinstance(prov, dict), "provenance sidecar present")
        if not isinstance(prov, dict):
            return
        sides = prov["inputs"]
        members = list(man.read_set)
        if flavor == "env":
            resolved = []
            for s in sides:
                d, _files = compare_env._find_input_dir(
                    Path(s["selection"]), row_key, "*.pdf")
                census = {m["name"]: (m["size"], m["mtime_ns"])
                          for m in s.get("members", [])}
                resolved.append((Path(d).resolve(), census))
            for m in members:
                mp = Path(m.name)
                parent = mp.resolve().parent
                match = next((c for d, c in resolved if parent == d), None)
                st = mp.stat() if mp.is_file() else None
                note(rec, match is not None,
                     f"read-set member under a compared side dir: {mp.name}")
                note(rec, st is not None and match is not None
                     and match.get(mp.name) == (st.st_size, st.st_mtime_ns),
                     f"read-set member in that side's census: {mp.name}")
        elif man.state == em.STATE_RENDERED:
            want = {str(Path(s["selection"]).resolve()): s["sha256"]
                    for s in sides}
            got = {str(Path(m.name).resolve()): m.sha256 for m in members}
            note(rec, got == want,
                 "read set == the comparison's two recorded documents "
                 f"(got {len(got)} member(s))")
        else:
            # A generation that rendered nothing (no differences to
            # illustrate) opened no source document, so its read set is
            # EMPTY by construction — requiring the two documents here would
            # demand a claim the generation never made. The state itself is
            # the assertion, and it is checked above via em.describe.
            note(rec, not members,
                 f"a non-rendered generation ({man.state}) claims no read "
                 f"set (got {len(members)} member(s))")
        if man.state != em.STATE_RENDERED:
            return
        wb_path = ve.sibling_paths(cmp_path)[0]
        src_lines, legends, rows, has_ledger = _summary_rows_of(wb_path)
        note(rec, has_ledger, "Ledger sheet present")
        want_lines = [f"Compared {s['role']}: {s['selection']}" for s in sides]
        note(rec, src_lines == want_lines,
             f"source lines are the provenance selections ({src_lines!r})")
        want_legend = ve._legend_for(
            ve.FLAVOR_ENV if flavor == "env" else ve.FLAVOR_TSN)
        note(rec, all(lg == want_legend for lg in legends),
             "image-sheet legends state the true sources")
        rec["examples"] = len(rows)
        # Panel-text fidelity census: values past the panel bound draw as a
        # visibly elided prefix (panel_cell_text) — legal under HF-05, so the
        # over-limit examples are LISTED for the native-scale image inspection
        # rather than failed; the drawn-string decision itself is unit-locked
        # in check_visual_evidence.
        rec["elided_examples"] = [
            [f, a] for f, a, va, vb in rows
            if len(va) > ve.PANEL_TEXT_MAX or len(vb) > ve.PANEL_TEXT_MAX]
        rec["values_over_26"] = sum(
            1 for _f, _a, va, vb in rows for v in (va, vb)
            if len(v) > PREFIX_PANEL_TRUNCATION_LIMIT)
        # PCOA-FINAL-005: panels whose compared value is blank on exactly ONE
        # side. The blank side has no glyph to box, so these are the hardest
        # targets to draw — recorded as a population WITH its identities, not
        # left to whichever ones a sample happened to include.
        rec["blank_side_examples"] = [
            [f, a, ("a" if not va else "b")] for f, a, va, vb in rows
            if bool(va) != bool(vb)]
        rec["blank_side_count"] = len(rec["blank_side_examples"])
        if flavor == "env":
            adapter = ve.env_adapter_for(row_key)
            rederived = 0
            for field, at, va, vb in rows:
                route, _, key = at.partition(" @ ")
                key = key or route
                vals = []
                boxes_ok = True
                for d, _census in resolved:
                    pdf = adapter.tsmis_pdf_path(d, route)
                    recs = adapter.env_locate(Path(pdf), {key})
                    got_recs = recs.get(key) or recs.get(route) or []
                    if len(got_recs) != 1:
                        boxes_ok = False
                        break
                    r0 = got_recs[0]
                    box = adapter.env_box(r0, field)
                    if box is None:
                        boxes_ok = False
                        break
                    _pg, cell, yspan, _xs = box
                    if cell[1] < yspan[0] - 3 or cell[3] > yspan[1] + 3:
                        boxes_ok = False
                        break
                    vals.append(adapter.env_value(r0, field))
                match = (boxes_ok and (vals[0] or "") == va
                         and (vals[1] or "") == vb)
                note(rec, match, f"env re-derivation matches ({field} @ {at})")
                rederived += bool(match)
            rec["env_rederived"] = rederived
        return

    # Driven by the DECLARATION, not by a glob: a set that was never produced
    # is validated as missing rather than skipped. `flavor` comes from the
    # placement too, so it no longer has to be re-guessed from the filename.
    for place in places:
        if place["identity"] not in required:
            continue
        validate_set(place["path"], place["flavor"],
                     row_key=place["row"] if place["flavor"] == "env" else None,
                     identity=place["identity"])
    # The silent lanes: nothing evidence-like anywhere under the PvE tree.
    pve_tree = side_root / "output" / "comparisons" / "pdf-vs-excel-by-day"
    stray = sorted(str(p) for p in pve_tree.rglob("*evidence*")) if pve_tree.is_dir() else []
    out["pve_stray"] = stray
    if stray:
        problems.append(f"PvE tree holds evidence artifacts: {stray}")
    out["problem_count"] = len(problems)
    write_json(root / "results" / f"validate-{side}.json", out)
    print(f"validate {side}: {len(out['sets'])}/{len(required)} declared "
          f"set(s), {len(problems)} problem(s)")
    if problems:
        for p in problems[:MAX_PROBLEMS_PRINTED]:
            print("  PROBLEM:", p)
        raise SystemExit(1)


def _com_rows(sheet):
    """Yield the sheet's A1-anchored rows as tuples, read in bounded chunks.

    Chunked rather than one whole-sheet `.Value`: the measured Highway Log
    comparison sheet is 63,710 rows x 41 columns, and materializing 2.6M cells
    in one COM call is a memory hazard. Chunking keeps the round trips few and
    the footprint bounded, and — unlike reading only the two located columns —
    it still yields the WHOLE row, which the all-blank-row rule needs.
    Normalizes the shapes COM returns for 1-cell and 1-row ranges."""
    used = sheet.UsedRange
    last_row = used.Row + used.Rows.Count - 1
    last_col = used.Column + used.Columns.Count - 1
    for start in range(1, last_row + 1, COM_ROW_CHUNK):
        stop = min(start + COM_ROW_CHUNK - 1, last_row)
        block = sheet.Range(sheet.Cells(start, 1),
                            sheet.Cells(stop, last_col)).Value
        if block is None:
            continue
        if not isinstance(block, tuple):
            yield (block,)
        elif block and not isinstance(block[0], tuple):
            yield block
        else:
            yield from block


def _com_comparison_counts(workbook):
    """(diff_cells, one_sided, data_rows) from an OPEN Excel workbook, or the
    unknown triple.

    The COM twin of `artifact_store.comparison_counts`, deliberately under the
    SAME contract: sheet "Comparison", and the Status/Diffs columns located by
    UNIQUE EXACT HEADER LABEL in row 1 — never by hard-coded position. It has
    to be COM rather than openpyxl because the FORMULAS twin holds formulas:
    only a live recalculation produces its numbers, and openpyxl reads cached
    values that a freshly built twin does not have."""
    try:
        sheet = workbook.Sheets(COMPARISON_SHEET)
    except Exception as e:  # noqa: BLE001 — no Comparison sheet is an answer
        print(f"  formulas twin has no '{COMPARISON_SHEET}' sheet: "
              f"{_first_line(e)}")
        return (None, None, None)
    rows_iter = _com_rows(sheet)
    header = next(rows_iter, None)
    if header is None:
        return (None, None, None)
    cols = {}
    for label in (STATUS_LABEL, DIFFS_LABEL):
        hits = [i + 1 for i, value in enumerate(header) if value == label]
        if len(hits) != 1:
            print(f"  formulas twin lacks a unique '{label}' header "
                  f"(found {len(hits)})")
            return (None, None, None)
        cols[label] = hits[0]
    diff_cells = one_sided = data_rows = 0
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        data_rows += 1
        status = row[cols[STATUS_LABEL] - 1]
        diffs = row[cols[DIFFS_LABEL] - 1]
        if status == BOTH_STATUS:
            if (isinstance(diffs, bool)
                    or not isinstance(diffs, (int, float))
                    or not float(diffs).is_integer() or diffs < 0):
                print(f"  formulas twin: matched row has invalid Diffs {diffs!r}")
                return (None, None, None)
            diff_cells += int(diffs)
        elif isinstance(status, str) and status:
            if diffs not in (None, ""):
                print("  formulas twin: one-sided row unexpectedly carries "
                      f"Diffs {diffs!r}")
                return (None, None, None)
            one_sided += 1
        else:
            print(f"  formulas twin: row has invalid Status {status!r}")
            return (None, None, None)
    return (diff_cells, one_sided, data_rows)


def _repair_logs_beside(folder):
    """Excel repair logs sitting next to a workbook. BEST EFFORT ONLY — see
    EXCEL_REPAIR_NOTE: Excel writes one in some configurations and not in
    others, so an empty result never proves a clean open."""
    try:
        return {p.resolve() for p in Path(folder).glob(EXCEL_REPAIR_LOG_GLOB)}
    except OSError as e:
        print(f"  repair-log sweep failed for {folder}: {_first_line(e)}")
        return set()


def _check_formulas_twins(excel, side_root, day_matrix, matrix, artifact_store,
                          problems):
    """Prove every placement's live-formulas twin SETTLES to the values
    workbook's own numbers. Returns one record per placement and appends any
    disagreement to `problems`.

    A comparison built with `also_formulas=True` publishes a values workbook
    (the numbers) and a `(formulas)` twin (the live recomputation). Nothing
    previously proved the twin agrees, so a twin that recalculates to different
    Status/Diffs totals — the whole reason it exists — would ship unnoticed.

    A twin is EXPECTED unless the PRODUCER itself declined to write one, and
    that decision is re-derived by calling the producer's own probe rather than
    re-inventing it: `matrix._comparison_row_count` reads openpyxl's stored
    dimension and returns None when the workbook stores none, in which case the
    producer writes the twin ANYWAY ("never skip the twin on an uncertain
    probe"). Re-deriving it from a scanned row count would have got this
    backwards — the measured Highway Log sheet scans to 63,710 rows, far over
    the limit, yet its twin exists because the probe returns None."""
    records = []
    for place in _placements(side_root, day_matrix):
        values = place["path"]
        rec = {"identity": place["identity"], "values": str(values),
               "formulas": None, "values_totals": None,
               "formulas_totals": None, "agree": None, "twin_expected": None,
               "note": ""}
        records.append(rec)
        if not values.is_file():
            rec["note"] = "the comparison workbook itself is missing"
            problems.append(
                f"{values}: no comparison workbook, so its live-formulas twin "
                "has nothing to settle against")
            continue
        diff_cells, one_sided, data_rows = artifact_store.comparison_counts(
            values)
        rec["values_totals"] = {"diff_cells": diff_cells,
                                "one_sided": one_sided, "rows": data_rows}
        if data_rows is None:
            rec["note"] = "the values workbook has no readable count contract"
            problems.append(
                f"{values}: unreadable Status/Diffs contract, so its twin "
                "cannot be settled against it")
            continue
        limit = matrix._FORMULAS_TWIN_MAX_ROWS
        probe = matrix._comparison_row_count(values)
        rec["producer_probe_rows"] = probe
        rec["twin_expected"] = probe is None or probe <= limit
        twin = Path(matrix._formulas_sibling(values))
        rec["formulas"] = str(twin)
        if not rec["twin_expected"]:
            rec["note"] = (f"the producer's own probe reads {probe} rows, over "
                           f"the {limit}-row limit, so it deliberately writes "
                           "no twin")
            continue
        if not twin.is_file():
            problems.append(
                f"{twin}: live-formulas twin missing, but the producer's own "
                f"probe ({probe} rows vs the {limit}-row limit) says one "
                "should have been written")
            continue
        print(f"  twin {twin.name}: recalculating...")
        book = None
        try:
            book = excel.Workbooks.Open(str(twin), UpdateLinks=0,
                                        ReadOnly=True, CorruptLoad=0)
            excel.CalculateFullRebuild()
            f_diff, f_one, f_rows = _com_comparison_counts(book)
        except Exception as e:  # noqa: BLE001 — recorded as this twin's problem
            f_diff = f_one = f_rows = None
            rec["error"] = f"{type(e).__name__}: {e}"
            problems.append(f"{twin}: recalculation failed — {rec['error']}")
        finally:
            if book is not None:
                book.Close(SaveChanges=False)
        rec["formulas_totals"] = {"diff_cells": f_diff, "one_sided": f_one,
                                  "rows": f_rows}
        rec["agree"] = (f_diff, f_one, f_rows) == (diff_cells, one_sided,
                                                   data_rows)
        if not rec["agree"]:
            problems.append(
                f"{twin}: the live-formulas twin settles to "
                f"diff_cells={f_diff} one_sided={f_one} rows={f_rows}, but the "
                f"values workbook says diff_cells={diff_cells} "
                f"one_sided={one_sided} rows={data_rows}")
        print(f"    twin agree={rec['agree']} "
              f"values={rec['values_totals']} formulas={rec['formulas_totals']}")
    return records


def phase_excel(root, side, tree, run_id):
    """Open every DECLARED evidence workbook in INSTALLED Excel: it must open,
    and it must still carry its Ledger sheet and its embedded images.

    WHAT THIS PHASE DOES NOT PROVE — REPAIR. It sets `DisplayAlerts = False`,
    which means that if Excel decides a workbook is damaged it repairs it
    SILENTLY and `Open()` still returns normally. Excel exposes no COM property
    saying "this file was repaired", so THIS PHASE CANNOT DETECT SILENT REPAIR.
    `opened` / `open_without_error` therefore mean exactly one thing — `Open()`
    did not raise — and must never be read, quoted, or summarised as "opened
    clean" or "zero repairs". `repair_logs` is a best-effort extra signal only.

    The population is derived from the manifests rather than from the glob: a
    set whose generation RENDERED owes a workbook, and a no_differences /
    no_examples generation owes none. A workbook found outside the declared
    placements is an extra and fails the phase.

    It ALSO proves formulas-twin settlement. Every placement is built with
    `also_formulas=True`, so each comparison has a values workbook and a
    live-formulas twin, and nothing previously proved the twin computes the
    same answer. Each twin is opened, fully recalculated, and its Status/Diffs
    totals compared against the values workbook's. NOTE: the twins are large
    (the Highway Log ones are hundreds of MB), so this phase is measured in
    tens of minutes, not seconds."""
    side_root = _sandbox(tree, root, side)
    import artifact_store
    import day_matrix
    import matrix
    import win32com.client  # type: ignore
    books = sorted((side_root / "store" / "comparisons").rglob("*(evidence).xlsx"))
    books += sorted((side_root / "output" / "comparisons").rglob("*(evidence).xlsx"))
    commit, dirty, stamp = runtime_stamp(root, tree, side)
    out = {"run_id": run_id, "side": side, "tree": str(tree),
           "tree_commit": commit, "tree_runtime_dirty": dirty,
           "runtime_stamp": stamp,
           "books": [], "problems": [], "repair_detection": EXCEL_REPAIR_NOTE}
    # Which placements owe a workbook, per their own published manifest state.
    required_books, states = {}, {}
    for place in _placements(side_root, day_matrix):
        state = _evidence_state(place["path"])
        states[place["identity"]] = state["manifest_state"]
        if state["manifest_state"] == EVIDENCE_STATE_RENDERED:
            required_books[place["identity"]] = _norm(
                _evidence_sibling_paths(place["path"])[0])
    pop = population_diff(required_books, [_norm(b) for b in books])
    out["manifest_states"] = states
    out["population"] = dict(pop, required=len(required_books),
                             discovered=len(books))
    for label, items in (("evidence WORKBOOK missing for a rendered set",
                          pop["missing"]),
                         ("UNEXPECTED evidence workbook", pop["extra"]),
                         ("DUPLICATE workbook identity", pop["duplicate"])):
        out["problems"].extend(f"{label}: {item}" for item in items)
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        for b in books:
            entry = {"path": str(b)}
            before = _repair_logs_beside(b.parent)
            try:
                wb = excel.Workbooks.Open(str(b), UpdateLinks=0, ReadOnly=True,
                                          CorruptLoad=0)
                sheets = [s.Name for s in wb.Sheets]
                entry["sheets"] = len(sheets)
                entry["has_ledger"] = LEDGER_SHEET in sheets
                entry["pictures"] = sum(
                    wb.Sheets(n).Shapes.Count for n in sheets
                    if n not in (SUMMARY_SHEET, LEDGER_SHEET))
                wb.Close(SaveChanges=False)
                entry["opened"] = True
                entry["open_without_error"] = True
            except Exception as e:  # noqa: BLE001
                entry["opened"] = False
                entry["open_without_error"] = False
                entry["error"] = f"{type(e).__name__}: {e}"
                out["problems"].append(f"{b}: Open() raised — {entry['error']}")
            entry["repair_logs"] = sorted(
                str(p) for p in _repair_logs_beside(b.parent) - before)
            if entry["repair_logs"]:
                out["problems"].append(
                    f"{b}: Excel wrote a repair log {entry['repair_logs']}")
            if entry["opened"]:
                if not entry["has_ledger"]:
                    out["problems"].append(
                        f"{b}: no '{LEDGER_SHEET}' sheet")
                if entry["pictures"] < MIN_EMBEDDED_PICTURES:
                    out["problems"].append(
                        f"{b}: {entry['pictures']} embedded picture(s) across "
                        f"its image sheets, expected at least "
                        f"{MIN_EMBEDDED_PICTURES}")
            out["books"].append(entry)
            print(f"  excel {b.name}: opened={entry['opened']} "
                  f"sheets={entry.get('sheets')} pics={entry.get('pictures')} "
                  f"ledger={entry.get('has_ledger')}")
        out["twins"] = _check_formulas_twins(
            excel, side_root, day_matrix, matrix, artifact_store,
            out["problems"])
    finally:
        excel.Quit()
    out["problem_count"] = len(out["problems"])
    write_json(root / "results" / f"excel-{side}.json", out)
    print(f"excel {side}: {len(books)}/{len(required_books)} declared "
          f"workbook(s), {len(out['problems'])} problem(s)")
    if out["problems"]:
        for p in out["problems"][:MAX_PROBLEMS_PRINTED]:
            print("  PROBLEM:", p)
        raise SystemExit(1)


def phase_checks_at_base(root, tree, head_tree):
    """Run the HEAD's extended evidence checks against the BASE scripts tree —
    the red half of red→green — and ASSERT the result of each one.

    This phase used to record exit codes and stop, which proved nothing: a
    check that crashed for an unrelated reason, or that died importing a
    function the hotfix adds and so never reached its own assertion, was
    indistinguishable from a check that genuinely detected the defect. Every
    check is now declared in BASE_CHECK_EXPECTATIONS as must-be-red (with the
    substring its output must carry) or as a must-be-green control, and each
    run is classified red / green / inconclusive by `classify_base_check`.

    The phase fails if ANY check's classification differs from its declaration
    — a must-be-red that came back inconclusive has not demonstrated red, and a
    control that came back red means the declaration is wrong. An
    `inconclusive` verdict is reported distinctly and is never counted as red;
    the usual remedy is to make the check load against the base runtime (probe
    for the new API and fail the ASSERTION) instead of dying on import."""
    import subprocess
    base_tree = Path(tree)
    # Derived from the contract so the run list and the declaration cannot drift.
    checks = tuple(BASE_CHECK_EXPECTATIONS)
    stage = root / "base-red"
    if stage.exists():
        shutil.rmtree(stage)
    stage.mkdir(parents=True)
    # The checks resolve scripts/ relative to their own location, so the stage
    # mirrors the tree: the BASE runtime (scripts + version.py) with only the
    # HEAD's check files overlaid into the base build/.
    shutil.copytree(base_tree / "scripts", stage / "scripts")
    shutil.copyfile(base_tree / "version.py", stage / "version.py")
    build_dir = stage / "build"
    shutil.copytree(base_tree / "build", build_dir)
    for name in checks + ("_checklib.py",):
        shutil.copyfile(Path(head_tree) / "build" / name, build_dir / name)
    # Both sides self-stamp, and each is asserted not to be the other: the
    # whole red/green claim is meaningless if the "base" runtime turns out to
    # be the acceptance head.
    base_commit, base_dirty, base_stamp = runtime_stamp(root, base_tree, "base")
    commit, dirty, head_stamp = runtime_stamp(root, head_tree, "head")
    out = {"base_tree": str(base_tree), "head_tree": str(head_tree),
           "head_tree_commit": commit, "head_tree_runtime_dirty": dirty,
           "base_tree_commit": base_commit,
           "base_tree_runtime_dirty": base_dirty,
           "runtime_stamp": {"base": base_stamp, "head": head_stamp},
           "results": {}}
    env = dict(os.environ, PYTHONIOENCODING="utf-8")
    problems = []
    for name in checks:
        spec = BASE_CHECK_EXPECTATIONS[name]
        expectation, signature = spec["expect"], spec["signature"]
        proc = subprocess.run(
            [sys.executable, str(build_dir / name)], cwd=str(stage),
            capture_output=True, text=True, encoding="utf-8",
            errors="replace", env=env)
        fails = [ln for ln in (proc.stdout or "").splitlines()
                 if "FAIL" in ln][:CHECK_FAIL_LINES_KEPT]
        tail = (proc.stdout or "").splitlines()[-CHECK_STDOUT_TAIL_LINES:]
        err_tail = (proc.stderr or "").splitlines()[-CHECK_STDERR_TAIL_LINES:]
        verdict = classify_base_check(proc.returncode, proc.stdout,
                                      proc.stderr, signature)
        classification = verdict["classification"]
        # The four `fail_lines` / `stdout_tail` / `stderr_tail` / `exit` keys
        # are the record's original shape and stay exactly as they were; the
        # verdict fields are added beside them.
        out["results"][name] = dict(
            verdict, exit=proc.returncode, fail_lines=fails,
            stdout_tail=tail, stderr_tail=err_tail,
            name=name, exit_code=proc.returncode,
            expected=expectation, signature=signature, why=spec["why"])
        print(f"  base-red {name}: exit={proc.returncode} "
              f"({len(fails)} FAIL line(s)) -> {classification} "
              f"[declared {expectation}] — {verdict['reason']}")
        if classification != expectation:
            problems.append(
                f"{name}: declared {expectation}, classified {classification} "
                f"— {verdict['reason']} (exit={proc.returncode}, "
                f"signature_found={verdict['signature_found']}, "
                f"signature={signature!r})")
        else:
            stale = stale_inconclusive_problem(name, expectation, fails)
            if stale:
                problems.append(stale)
    out["problems"] = problems
    out["problem_count"] = len(problems)
    out["classification_counts"] = {
        verdict: sum(1 for r in out["results"].values()
                     if r["classification"] == verdict)
        for verdict in (CHECK_RED, CHECK_GREEN, CHECK_INCONCLUSIVE)}
    # Named, not just counted: this is the bundle's ACTUAL red-at-base proof,
    # and a reader must be able to see how narrow it is without re-deriving it.
    out["red_evidence_checks"] = sorted(
        n for n, r in out["results"].items() if r["classification"] == CHECK_RED)
    if len(out["red_evidence_checks"]) < MIN_RED_AT_BASE_CHECKS:
        problems.append(
            f"no check demonstrated the defect at base ({len(checks)} ran, "
            f"at least {MIN_RED_AT_BASE_CHECKS} must be red) — the phase "
            "proves only that the checks execute")
        out["problem_count"] = len(problems)
    write_json(root / "results" / "base-red-checks.json", out)
    print(f"base-red: {out['classification_counts']}, "
          f"{len(problems)} problem(s)")
    print(f"  red-at-base evidence comes from {len(out['red_evidence_checks'])}"
          f" of {len(checks)} check(s): {out['red_evidence_checks']}")
    if problems:
        for p in problems[:MAX_PROBLEMS_PRINTED]:
            print("  PROBLEM:", p)
        raise SystemExit(1)


def phase_census(root):
    out = {"audit_everything": _census_one_root(AUDIT_EVERYTHING,
                                                "audit-everything"),
           "audit_byday": _census_one_root(AUDIT_BYDAY, "audit-byday"),
           "base_everything": [], "base_byday": []}
    base_tsn = root / "base" / "store" / "comparisons" / "tsn"
    if base_tsn.is_dir():
        out["base_everything"] = _census_one_root(base_tsn, "base-everything")
    base_byday = (root / "base" / "output" / "comparisons" / "tsn-by-day"
                  / f"{BYDAY_DAY} {BYDAY_SOURCE}")
    if base_byday.is_dir():
        out["base_byday"] = _census_one_root(base_byday, "base-byday")
    for key in list(out):
        sets = out[key]
        examined = sum(len(s["examples"]) for s in sets)
        truncated = sum(len(s["truncated_examples"]) for s in sets)
        elided = sum(len(s["legally_elided_examples"]) for s in sets)
        out[f"{key}_totals"] = {
            "sets": len(sets),
            "examples": examined,
            "truncated": truncated,
            "sets_with_pdf_read_members": sum(
                1 for s in sets if s["read_set_pdf"]),
            "blank_side_targets": sum(
                len(s["blank_side_examples"]) for s in sets),
            # The recount stated as a POPULATION, so the record can report a
            # rate instead of a bare count. `legally_elided` is a SUBSET of
            # `truncated_over_prefix_limit` (the elision bound is far above the
            # pre-fix cut) and is NOT a defect — the pre-fix defect population
            # is `truncated_not_elided`.
            "truncation_population": {
                "examples_examined": examined,
                "truncated_over_prefix_limit": truncated,
                "legally_elided": elided,
                "truncated_not_elided": truncated - elided,
                "prefix_limit": PREFIX_PANEL_TRUNCATION_LIMIT,
                "elision_limit": PANEL_TEXT_MAX_DECLARED,
            },
        }
    write_json(root / "results" / "prefix-defect-census.json", out)


# --------------------------------------------------------------------------- #
def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--run-id", default=RUN_ID_DEFAULT)
    ap.add_argument("--phase", required=True,
                    choices=("provision", "generate", "cameras", "census",
                             "checks-at-base", "validate", "excel", "counts"))
    ap.add_argument("--side", choices=("base", "head"),
                    help="generate: which acceptance side to run")
    ap.add_argument("--tree", default=str(REPO),
                    help="generate: the scripts tree under test")
    ap.add_argument("--root", default=str(ROOT_DEFAULT))
    ap.add_argument("--cells", default="",
                    help="generate: comma list of cell kinds to run "
                         "(everything-tsn,everything-self,everything-env,"
                         "byday,pve,classic); empty = all. `classic` reuses the "
                         "everything-tsn/-env placements' own inputs, so those "
                         "cells must already exist")
    ap.add_argument("--label", default="",
                    help="generate: suffix for the results/log files so a "
                         "partial re-run never clobbers a prior record")
    args = ap.parse_args(argv)
    assert_declared_counts()
    root = Path(args.root)
    root.mkdir(parents=True, exist_ok=True)
    sys.stdout = sys.stderr = Tee(root / "logs" / f"driver-{args.phase}"
                                  f"{('-' + args.side) if args.side else ''}.log")
    print(f"— run_rb4_acceptance {args.run_id} · phase={args.phase}"
          f" side={args.side} tree={args.tree} at {time.strftime('%F %T')}")
    if args.phase == "provision":
        phase_provision(root, ("base", "head"), Path(args.tree))
    elif args.phase == "generate":
        if not args.side:
            ap.error("--side is required for generate")
        kinds = tuple(k for k in args.cells.split(",") if k) or None
        phase_generate(root, args.side, Path(args.tree), args.run_id,
                       kinds=kinds, label=args.label)
    elif args.phase == "cameras":
        if not args.side:
            ap.error("--side is required for cameras")
        phase_cameras(root, args.side, Path(args.tree), args.run_id)
    elif args.phase == "census":
        phase_census(root)
    elif args.phase == "checks-at-base":
        phase_checks_at_base(root, Path(args.tree), REPO)
    elif args.phase == "counts":
        if not args.side:
            ap.error("--side is required for counts")
        phase_counts(root, args.side, Path(args.tree), args.run_id)
    elif args.phase == "validate":
        if not args.side:
            ap.error("--side is required for validate")
        phase_validate(root, args.side, Path(args.tree), args.run_id)
    elif args.phase == "excel":
        if not args.side:
            ap.error("--side is required for excel")
        phase_excel(root, args.side, Path(args.tree), args.run_id)
    print(f"— phase {args.phase} complete at {time.strftime('%F %T')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
