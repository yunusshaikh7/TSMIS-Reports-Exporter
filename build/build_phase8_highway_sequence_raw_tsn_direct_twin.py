#!/usr/bin/env python3
"""Build a direct-source raw-TSN audit twin for Highway Sequence.

This audit-only lane reads each of the twelve authoritative district PDFs into
one immutable byte payload, validates that exact payload against the accepted
Stage-6 bindings, and passes the same captured bytes to the independent
accepted Stage-6 parser.  It emits a product-consumable eight-column workbook
for all 69,804 records, including the 46 blank-County EQUATES annotations, plus
exhaustive row provenance and deterministic manifests.

The output is an input fixture for later product/evidence certification.  It is
not a Stage-8 family acceptance artifact and cannot make the family accepted.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
import hashlib
import importlib.metadata
import io
import json
import os
from pathlib import Path
import stat as stat_module
import sys
import tempfile
import time
from typing import Iterable, Mapping, Sequence
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill


BUILD_ROOT = Path(__file__).resolve().parent
REPO_ROOT = BUILD_ROOT.parent
sys.path.insert(0, str(BUILD_ROOT))

# Accepted audit-owned parser.  No product parser, comparator, schema, or
# evidence adapter is imported anywhere in this lane.
import phase6_highway_sequence_conservation as stage6  # noqa: E402


RAW_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library"
    r"\highway_sequence\raw"
)
VISUAL_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd"
)
STAGE6_RESULT = (
    VISUAL_ROOT / "phase6_tsn_conservation"
    / "highway_sequence_conservation_r7.json"
)
STAGE6_DECISION = Path(str(STAGE6_RESULT) + ".acceptance.json")
NORMALIZED_TSN = (
    VISUAL_ROOT / "phase4_tsn_rebaseline" / "raw-2026-07-12-r7"
    / "highway_sequence" / "consolidated"
    / "tsn_highway_sequence_normalized.xlsx"
)

WORKBOOK_NAME = "highway_sequence_raw_tsn_audit_twin.xlsx"
PROVENANCE_NAME = "highway_sequence_raw_tsn_audit_twin.provenance.json"
MANIFEST_NAME = "manifest.json"
RESULT_NAME = "result.json"
OUTPUT_NAMES = (
    MANIFEST_NAME, PROVENANCE_NAME, RESULT_NAME, WORKBOOK_NAME,
)

SHEET_NAME = "Highway Locations (TSN)"
HEADERS = (
    "Route", "County", "PM", "City", "HG", "FT",
    "Distance To Next Point", "Description",
)
FIXED_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
FIXED_DOCUMENT_TIME = datetime(2000, 1, 1, 0, 0, 0)
FIXED_CORE_TIME = "2000-01-01T00:00:00Z"
CORE_NAMESPACES = {
    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dc": "http://purl.org/dc/elements/1.1/",
    "dcterms": "http://purl.org/dc/terms/",
    "dcmitype": "http://purl.org/dc/dcmitype/",
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
}
EXPECTED_ORDERED_ROWS_SHA256 = (
    "5ef81b31622730e8f1369d1989cc92c717be7eb4ad8f29061b3750ff78f767fc"
)

EXPECTED_RAW_BINDINGS = (
    ("D01 HSL TSN.pdf", 204_709, "3a4cb30340a55edae2f72d758dcda62d30e21d919ecc862ec6955d6795252a4a"),
    ("D02 HSL TSN.pdf", 288_696, "f32078eb79f38fa2e4799319bd10f661ecdff669dd7c4ade18a5326723ad5d85"),
    ("D03 HSL TSN.pdf", 373_387, "8c5cd4638dd4901797f9c15e6fac7f998d5bc989749f874e6eedf52f72506fb0"),
    ("D04 HSL TSN.pdf", 625_052, "5facc297fd7d28e8ad760cce8d7f4699b1ee4bc7582f2a007196c0bf739bcd5a"),
    ("D05 HSL TSN.pdf", 265_876, "b8246f8c28e31d0c4acc352b7148988b6a6a0d7abaf56e810943e14816389e7b"),
    ("D06 HSL TSN.pdf", 327_246, "e240f038390109ca02ceb012a5e8e5b82fc8845c49be718506acb56667db3dad"),
    ("D07 HSL TSN.pdf", 555_648, "c791b99789e496efb83b52850aa54e142946aaa541a91b780489fe7e0bc7ec25"),
    ("D08 HSL TSN.pdf", 370_505, "f23b8e3d5a90200cc1a6285ebb40480b828673f9e5a37b06f36fe30bc9697565"),
    ("D09 HSL TSN.pdf", 103_868, "c6984a7e947ff600a450e4387f318aeed4826b05249361a694fbe507d0c7c5c3"),
    ("D10 HSL TSN.pdf", 298_313, "e510a575c56c5af4404968d9fe51271f79cc23377df1e5c651b45b563dbf2ed6"),
    ("D11 HSL TSN.pdf", 315_238, "920e3e352c1f24be415271c9819fc8bddce8ac6ef3095684e9fe06c87cf7378b"),
    ("D12 HSL TSN.pdf", 138_411, "5583c0a0b94feeddaefda8bfa35bf34657cfb9f3b8e0a8d2b047c8fc27cbcc7a"),
)
EXPECTED_NON_SOURCE = (
    "_PUT TSN FILES HERE.txt", 456,
    "9048d67bfd2e3fd7515f95e849e33479bd602669eec35b567b77c0cfbef51eda",
)
STATIC_BINDINGS = {
    "stage6_oracle": {
        "path": BUILD_ROOT / "phase6_highway_sequence_conservation.py",
        "bytes": 63_233,
        "sha256": "0d6cacfa5a4615a80381b077780b051127958bbf325979cf24b7a5c29eb8e17b",
    },
    "xlsx_reader": {
        "path": BUILD_ROOT / "phase3_xlsx_stream.py",
        "bytes": 40_888,
        "sha256": "bbfda5ccdbea3697978c0ba4414b7dccf3d5c248ba6762aa946c76e920fc940b",
    },
    "accepted_stage6_result": {
        "path": STAGE6_RESULT,
        "bytes": 1_276_684,
        "sha256": "bdd344258ced0e138196c518be2d49ee058f5f9c0f52dea860c328fc3216d1e2",
    },
    "accepted_stage6_decision": {
        "path": STAGE6_DECISION,
        "bytes": 5_934,
        "sha256": "71fe59a5f4676d3b935bcbea380374b14fdccfd77b674ea88148fa18760ffde2",
    },
    "accepted_normalized_tsn": {
        "path": NORMALIZED_TSN,
        "bytes": 2_536_901,
        "sha256": "9dc84c661a9284131baf928767e210a6d708c0a338819fca2b69b907f85dd041",
    },
}


class DirectTwinError(RuntimeError):
    """A source binding, topology, workbook, or lifecycle invariant failed."""


@dataclass(frozen=True)
class CapturedPayload:
    name: str
    canonical_path: str
    bytes: int
    sha256: str
    stat_signature: tuple[int, int, int, int]
    payload: bytes

    def public_identity(self) -> dict[str, object]:
        return {
            "name": self.name,
            "canonical_path": self.canonical_path,
            "bytes": self.bytes,
            "sha256": self.sha256,
        }


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise DirectTwinError(message)


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _json_bytes(value: object) -> bytes:
    return (
        json.dumps(
            value, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
        )
        + "\n"
    ).encode("utf-8")


def _artifact_identity(payload: bytes) -> dict[str, object]:
    return {"bytes": len(payload), "sha256": _sha_bytes(payload)}


def _stat_signature(path: Path) -> tuple[int, int, int, int]:
    stat = path.stat()
    return stat.st_dev, stat.st_ino, stat.st_size, stat.st_mtime_ns


def _capture_unbound_once(path: Path) -> CapturedPayload:
    _require_plain_path(path.parent, f"parent of {path.name}")
    _require_plain_path(path, path.name)
    _require(path.is_file(), f"not a plain source file: {path}")
    before = _stat_signature(path)
    payload = path.read_bytes()
    after = _stat_signature(path)
    _require(before == after, f"source changed during its single-byte capture: {path}")
    observed_sha256 = _sha_bytes(payload)
    return CapturedPayload(
        name=path.name,
        canonical_path=str(path.resolve()),
        bytes=len(payload),
        sha256=observed_sha256,
        stat_signature=after,
        payload=payload,
    )


def _capture_once(path: Path, expected_bytes: int, expected_sha256: str) -> CapturedPayload:
    captured = _capture_unbound_once(path)
    _require(
        captured.bytes == expected_bytes and captured.sha256 == expected_sha256,
        f"source identity drift: {path}: {captured.bytes}/{captured.sha256}",
    )
    return captured


def _capture_static_inputs() -> tuple[dict[str, object], dict[str, bytes]]:
    identities: dict[str, object] = {}
    payloads: dict[str, bytes] = {}
    for label, binding in STATIC_BINDINGS.items():
        captured = _capture_once(
            Path(binding["path"]), int(binding["bytes"]), str(binding["sha256"]),
        )
        identities[label] = captured.public_identity()
        payloads[label] = captured.payload
    return identities, payloads


def _is_reparse_point(path: Path) -> bool:
    observed = os.lstat(path)
    attributes = int(getattr(observed, "st_file_attributes", 0))
    reparse_flag = int(getattr(stat_module, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400))
    return path.is_symlink() or bool(attributes & reparse_flag)


def _require_plain_components(path: Path, label: str) -> None:
    """Reject a reparse point in every existing lexical component.

    Resolution is intentionally forbidden until after this walk: resolving
    first would erase the alias component the audit is meant to detect.
    """
    absolute = path if path.is_absolute() else Path(os.path.abspath(path))
    _require(absolute.is_absolute(), f"{label} is not absolute: {path}")
    parts = absolute.parts
    _require(bool(parts), f"{label} has no lexical path components")
    current = Path(parts[0])
    components_checked = 0
    if os.path.lexists(current):
        _require(not _is_reparse_point(current), f"{label} drive/root is a reparse point: {current}")
        components_checked += 1
    for part in parts[1:]:
        current = current / part
        if not os.path.lexists(current):
            break
        _require(
            not _is_reparse_point(current),
            f"{label} contains a reparse-point component: {current}",
        )
        components_checked += 1
    _require(components_checked > 0, f"{label} has no existing lexical component")


def _require_plain_path(path: Path, label: str) -> None:
    _require_plain_components(path, label)
    _require(path.exists(), f"{label} is missing: {path}")
    _require(not _is_reparse_point(path), f"{label} is a reparse point/alias: {path}")


def _paths_overlap(left: Path, right: Path) -> bool:
    left = left.resolve(strict=False)
    right = right.resolve(strict=False)
    return left == right or left in right.parents or right in left.parents


def _output_policy_violations(candidate: Path) -> list[str]:
    violations: list[str] = []
    if not candidate.is_absolute():
        violations.append("not_absolute")
    if any(part in (".", "..") for part in candidate.parts):
        violations.append("lexical_alias_component")
    canonical = candidate.resolve(strict=False)
    visual = VISUAL_ROOT.resolve()
    if canonical == visual or visual not in canonical.parents:
        violations.append("outside_private_visual_root")
    if _paths_overlap(canonical, RAW_ROOT):
        violations.append("overlaps_authoritative_raw_root")
    for label, binding in STATIC_BINDINGS.items():
        artifact = Path(binding["path"]).resolve()
        if _paths_overlap(canonical, artifact):
            violations.append(f"overlaps_static_artifact:{label}")
        if _paths_overlap(canonical, artifact.parent):
            violations.append(f"overlaps_static_parent_tree:{label}")
    return sorted(set(violations))


def _reparse_component_control() -> dict[str, object]:
    control_root = Path(tempfile.mkdtemp(
        prefix=".direct-twin-reparse-control-", dir=VISUAL_ROOT,
    ))
    real_root = control_root / "real"
    real_nested = real_root / "level-one"
    alias = control_root / "alias"
    rejected = False
    try:
        real_nested.mkdir(parents=True, exist_ok=False)
        os.symlink(real_root, alias, target_is_directory=True)
        _require(_is_reparse_point(alias), "planted directory alias is not a reparse point")
        candidate = alias / "level-one" / "level-two" / "output"
        try:
            _require_plain_components(candidate, "planted multi-level alias control")
        except DirectTwinError as error:
            rejected = "reparse-point component" in str(error)
        _require(rejected, "multi-level reparse component escaped lexical walk")
        return {
            "kind": "directory_symlink_reparse_point",
            "alias_is_multiple_components_above_output": True,
            "rejected_before_resolve": True,
        }
    finally:
        if os.path.lexists(alias):
            alias.unlink()
        if real_nested.exists():
            real_nested.rmdir()
        if real_root.exists():
            real_root.rmdir()
        if control_root.exists():
            control_root.rmdir()


def _output_gate_controls() -> dict[str, object]:
    valid = VISUAL_ROOT / "__direct_twin_disposable_valid_child__"
    raw_child = RAW_ROOT / "__direct_twin_disposable_forbidden_child__"
    static_child = STAGE6_RESULT.parent / "__direct_twin_disposable_forbidden_child__"
    lexical_alias = (
        VISUAL_ROOT / "phase6_tsn_conservation" / ".."
        / "__direct_twin_disposable_alias_child__"
    )
    valid_violations = _output_policy_violations(valid)
    raw_violations = _output_policy_violations(raw_child)
    static_violations = _output_policy_violations(static_child)
    alias_violations = _output_policy_violations(lexical_alias)
    _require(not valid_violations, f"valid output control rejected: {valid_violations}")
    _require(
        "overlaps_authoritative_raw_root" in raw_violations,
        f"raw-child output control escaped: {raw_violations}",
    )
    _require(
        any(item.startswith("overlaps_static_parent_tree:") for item in static_violations),
        f"static-child output control escaped: {static_violations}",
    )
    _require(
        "lexical_alias_component" in alias_violations,
        f"lexical-alias output control escaped: {alias_violations}",
    )
    reparse_control = _reparse_component_control()
    return {
        "valid_private_child_accepted": True,
        "raw_root_child_rejected": True,
        "static_artifact_parent_child_rejected": True,
        "lexical_parent_alias_rejected": True,
        "multi_level_reparse_component_rejected": reparse_control[
            "rejected_before_resolve"
        ],
        "reparse_control": reparse_control,
    }


def _validate_output_root(output_root: Path) -> tuple[Path, dict[str, object]]:
    # Walk every supplied/existing lexical component before the first resolve.
    # This ordering is part of the security contract: resolving first would
    # erase the exact reparse component the gate must reject.
    _require_plain_components(output_root, "supplied output root")
    _require_plain_components(VISUAL_ROOT, "private visual root")
    _require_plain_components(RAW_ROOT, "authoritative raw root")
    for label, binding in STATIC_BINDINGS.items():
        artifact = Path(binding["path"])
        _require_plain_components(artifact.parent, f"{label} parent")
        _require_plain_components(artifact, label)
    violations = _output_policy_violations(output_root)
    _require(not violations, f"unsafe output root policy: {violations}")
    canonical = output_root.resolve(strict=False)
    _require(not output_root.exists(), f"refusing to overwrite output root: {output_root}")
    _require(output_root.parent.is_dir(), f"output parent does not exist: {output_root.parent}")
    _require_plain_path(VISUAL_ROOT, "private visual root")
    _require_plain_path(output_root.parent, "output parent")
    for label, binding in STATIC_BINDINGS.items():
        artifact = Path(binding["path"])
        _require_plain_path(artifact.parent, f"{label} parent")
        _require_plain_path(artifact, label)
    return canonical, _output_gate_controls()


def _capture_source_universe() -> tuple[list[CapturedPayload], CapturedPayload, dict[str, object]]:
    _require_plain_path(RAW_ROOT.parent, "authoritative raw parent")
    _require_plain_path(RAW_ROOT, "authoritative raw root")
    _require(RAW_ROOT.resolve() == Path(stage6.RAW_DIR).resolve(), "Stage-6 raw-root role drift")
    _require(tuple(stage6.RAW_BINDINGS) == EXPECTED_RAW_BINDINGS, "Stage-6 raw bindings drift")
    _require(tuple(stage6.NON_SOURCE_NAMES) == (EXPECTED_NON_SOURCE[0],), "Stage-6 non-source role drift")

    entries = sorted(RAW_ROOT.iterdir(), key=lambda path: path.name)
    expected_names = sorted([
        *(name for name, _size, _digest in EXPECTED_RAW_BINDINGS),
        EXPECTED_NON_SOURCE[0],
    ])
    _require([path.name for path in entries] == expected_names, "raw file universe drift")
    for entry in entries:
        _require_plain_path(entry, f"raw-universe entry {entry.name}")
        _require(entry.is_file(), f"raw universe contains a non-file: {entry}")

    captured = [
        _capture_once(RAW_ROOT / name, size, digest)
        for name, size, digest in EXPECTED_RAW_BINDINGS
    ]
    non_source = _capture_once(
        RAW_ROOT / EXPECTED_NON_SOURCE[0],
        EXPECTED_NON_SOURCE[1], EXPECTED_NON_SOURCE[2],
    )
    _require(sum(item.bytes for item in captured) == 3_866_949, "raw source byte-total drift")
    identities = [item.public_identity() for item in captured]
    return captured, non_source, {
        "source_members": len(captured),
        "source_bytes": sum(item.bytes for item in captured),
        "source_identities": identities,
        "source_identity_ledger_sha256": _sha_bytes(_json_bytes(identities)),
        "non_source_role": non_source.public_identity(),
        "file_universe": expected_names,
        "each_payload_captured_once": True,
        "parser_consumes_the_same_captured_payload_object": True,
    }


def _validate_stage6_chain(
    static_identities: Mapping[str, object],
    static_payloads: Mapping[str, bytes],
    source_capture: Mapping[str, object],
) -> tuple[dict[str, object], dict[str, object]]:
    result = json.loads(static_payloads["accepted_stage6_result"])
    decision = json.loads(static_payloads["accepted_stage6_decision"])
    _require(isinstance(result, dict) and isinstance(decision, dict), "Stage-6 JSON shape drift")
    _require(
        result.get("stage6_family_audit_complete") is True
        and result.get("unexplained_projection_residue_count") == 0,
        "accepted Stage-6 terminal flags drift",
    )
    tracked = decision.get("tracked_identities", {})
    _require(
        decision.get("decision") == "accepted_stage6_family_audit"
        and tracked.get("result", {}).get("sha256")
        == static_identities["accepted_stage6_result"]["sha256"]
        and tracked.get("oracle", {}).get("sha256")
        == static_identities["stage6_oracle"]["sha256"]
        and tracked.get("reader", {}).get("sha256")
        == static_identities["xlsx_reader"]["sha256"]
        and tracked.get("normalized", {}).get("sha256")
        == static_identities["accepted_normalized_tsn"]["sha256"],
        "accepted Stage-6 detached decision chain drift",
    )

    decision_members = [
        {
            "name": Path(item["canonical_path"]).name,
            "canonical_path": str(Path(item["canonical_path"]).resolve()),
            "bytes": item["size"],
            "sha256": item["sha256"],
        }
        for item in decision.get("raw_member_identities", [])
    ]
    _require(
        decision_members == source_capture["source_identities"],
        "direct captured members differ from the accepted Stage-6 decision",
    )
    decision_non_source = decision.get("non_source_role_identities", [])
    _require(len(decision_non_source) == 1, "Stage-6 non-source decision census drift")
    accepted_non_source = decision_non_source[0]
    observed_non_source = source_capture["non_source_role"]
    _require(
        Path(accepted_non_source["canonical_path"]).resolve()
        == Path(observed_non_source["canonical_path"]).resolve()
        and accepted_non_source["size"] == observed_non_source["bytes"]
        and accepted_non_source["sha256"] == observed_non_source["sha256"],
        "Stage-6 non-source decision identity drift",
    )
    return result, {
        "decision": decision["decision"],
        "required_result_flags": decision["required_result_flags"],
        "tracked_identities": decision["tracked_identities"],
        "raw_member_identities_exact": True,
        "non_source_role_identity_exact": True,
    }


def _record_ref(record: Mapping[str, object]) -> dict[str, object]:
    return {
        "member": record["member"],
        "district": record["district"],
        "route": record["route"],
        "direction": record["direction"],
        "county": record["county"],
        "pm": record["pm"],
        "physical_page": record["physical_page"],
        "printed_page": record["printed_page"],
        "line": record["line"],
        "top": record["top"],
    }


def _owned_pair(annotation: Mapping[str, object], following: Mapping[str, object]) -> bool:
    return all(
        annotation[key] == following[key]
        for key in ("member", "district", "route", "direction")
    )


def _event_topology(
    document_records: Sequence[Sequence[Mapping[str, object]]],
) -> dict[str, object]:
    forward_pairs: list[dict[str, object]] = []
    reverse_pairs: list[dict[str, object]] = []
    forward_unpaired: list[dict[str, object]] = []
    reverse_unpaired: list[dict[str, object]] = []
    equate_rows = data_e_rows = 0

    for records in document_records:
        for index, record in enumerate(records):
            is_equate = record["kind"] == "equate"
            is_data_e = (
                record["kind"] == "data" and str(record["pm"]).endswith("E")
            )
            if is_equate:
                equate_rows += 1
                following = records[index + 1] if index + 1 < len(records) else None
                if (
                    following is None
                    or following["kind"] != "data"
                    or not str(following["pm"]).endswith("E")
                    or not _owned_pair(record, following)
                ):
                    forward_unpaired.append(_record_ref(record))
                else:
                    forward_pairs.append({
                        "annotation": _record_ref(record),
                        "data_e": _record_ref(following),
                    })
            if is_data_e:
                data_e_rows += 1
                preceding = records[index - 1] if index else None
                if (
                    preceding is None
                    or preceding["kind"] != "equate"
                    or not _owned_pair(preceding, record)
                ):
                    reverse_unpaired.append(_record_ref(record))
                else:
                    reverse_pairs.append({
                        "annotation": _record_ref(preceding),
                        "data_e": _record_ref(record),
                    })
    return {
        "equate_rows": equate_rows,
        "data_e_rows": data_e_rows,
        "forward_paired": len(forward_pairs),
        "reverse_paired": len(reverse_pairs),
        "forward_unpaired": len(forward_unpaired),
        "reverse_unpaired": len(reverse_unpaired),
        "forward_ledger_sha256": _sha_bytes(_json_bytes(forward_pairs)),
        "reverse_ledger_sha256": _sha_bytes(_json_bytes(reverse_pairs)),
        "pair_ledgers_exact": forward_pairs == reverse_pairs,
        "forward_unpaired_records": forward_unpaired,
        "reverse_unpaired_records": reverse_unpaired,
        "pairs": forward_pairs,
    }


def _forward_contract(topology: Mapping[str, object]) -> dict[str, object]:
    return {
        "equate_rows": topology["equate_rows"],
        "paired": topology["forward_paired"],
        "unpaired": topology["forward_unpaired"],
        "ledger_sha256": topology["forward_ledger_sha256"],
    }


def _reverse_contract(topology: Mapping[str, object]) -> dict[str, object]:
    return {
        "data_e_rows": topology["data_e_rows"],
        "paired": topology["reverse_paired"],
        "unpaired": topology["reverse_unpaired"],
        "ledger_sha256": topology["reverse_ledger_sha256"],
    }


def _reverse_topology_mutation(
    document_records: Sequence[Sequence[Mapping[str, object]]],
    baseline: Mapping[str, object],
) -> dict[str, object]:
    mutated = [[dict(record) for record in records] for records in document_records]
    selected: tuple[int, int] | None = None
    for document_index, records in enumerate(mutated):
        for index, record in enumerate(records):
            if (
                record["kind"] == "data"
                and not str(record["pm"]).endswith("E")
                and (index == 0 or records[index - 1]["kind"] != "equate")
            ):
                selected = document_index, index
                break
        if selected is not None:
            break
    _require(selected is not None, "no reverse-topology mutation candidate")
    document_index, row_index = selected
    record = mutated[document_index][row_index]
    before_pm = str(record["pm"])
    record["pm"] = before_pm + "E"

    observed = _event_topology(mutated)
    expected_forward = _forward_contract(baseline)
    expected_reverse = _reverse_contract(baseline)
    observed_forward = _forward_contract(observed)
    observed_reverse = _reverse_contract(observed)
    _require(
        observed_forward == expected_forward,
        "reverse-only mutation unexpectedly changed the forward topology contract",
    )
    _require(
        observed_reverse != expected_reverse
        and observed_reverse["data_e_rows"] == 999
        and observed_reverse["paired"] == 998
        and observed_reverse["unpaired"] == 1,
        "reverse-only mutation escaped the reverse topology gate",
    )
    return {
        "label": "unpaired data-E row must be rejected by reverse topology",
        "rejected": True,
        "mutation": {
            "source_ref": _record_ref(record),
            "before_pm": before_pm,
            "after_pm": record["pm"],
        },
        "forward_contract_unchanged": True,
        "reverse_contract_changed": True,
        "expected_forward": expected_forward,
        "observed_forward": observed_forward,
        "expected_reverse": expected_reverse,
        "observed_reverse": observed_reverse,
    }


def _parse_captured_sources(
    captured: Sequence[CapturedPayload],
) -> tuple[
    list[dict[str, object]], list[dict[str, object]],
    list[list[dict[str, object]]], dict[int, dict[str, int]],
]:
    all_records: list[dict[str, object]] = []
    documents: list[dict[str, object]] = []
    document_records: list[list[dict[str, object]]] = []
    source_ordinals: dict[int, dict[str, int]] = {}
    global_ordinal = 0
    for document_ordinal, member in enumerate(captured, 1):
        records, document = stage6._parse_document(member.name, member.payload)
        for member_ordinal, record in enumerate(records, 1):
            global_ordinal += 1
            source_ordinals[id(record)] = {
                "document_ordinal": document_ordinal,
                "member_record_ordinal": member_ordinal,
                "global_parse_ordinal": global_ordinal,
            }
        all_records.extend(records)
        documents.append(document)
        document_records.append(records)
    sorted_records = stage6._sorted_source(all_records)
    return sorted_records, documents, document_records, source_ordinals


def _fresh_stage6_digests(
    records: Sequence[Mapping[str, object]],
    documents: Sequence[Mapping[str, object]],
    accepted: Mapping[str, object],
) -> dict[str, object]:
    source_rows = [stage6._source_row(record) for record in records]
    provenance_headers = (
        "Member", "Physical Page", "Printed Page", "Line", "Top", "Raw Text",
    )
    provenance_rows = [(
        record["member"], record["physical_page"], record["printed_page"],
        record["line"], record["top"], record["raw_text"],
    ) for record in records]
    metadata_headers = (
        "Member", "District", "Report ID", "Report Title",
        "Cover Reference Date", "Data Reference Date", "Generation Time",
        "Policy SHA256", "PDF Metadata",
    )
    metadata_rows = []
    for document in documents:
        generation_times = sorted({
            claim["generation_time"] for claim in document["data_page_claims"]
        })
        metadata_rows.append((
            document["member"], document["district"], "OTM22025",
            "Highway Locations", "15-SEP-25", "15 SEP 2025",
            "|".join(generation_times), document["policy_sha256"],
            json.dumps(
                document["pdf_metadata"], sort_keys=True, separators=(",", ":"),
            ),
        ))
    fresh = {
        "raw_source_digests": stage6._dataset_digests(stage6.SOURCE_HEADERS, source_rows),
        "raw_source_provenance_digests": stage6._dataset_digests(provenance_headers, provenance_rows),
        "document_metadata_digests": stage6._dataset_digests(metadata_headers, metadata_rows),
    }
    for label, observed in fresh.items():
        _require(observed == accepted[label], f"fresh direct-source {label} drift")
    return {**fresh, "all_equal_accepted_stage6": True}


def _row_values(record: Mapping[str, object]) -> tuple[object, ...]:
    return (
        record["route"], record["county"], record["pm"], record["city"],
        record["hg"], record["ft"], record["distance"], record["description"],
    )


def _ordered_digest(rows: Iterable[Sequence[object]]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update(json.dumps(
            list(row), ensure_ascii=False, separators=(",", ":"),
        ).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def _validate_rows(
    records: Sequence[Mapping[str, object]],
) -> tuple[list[tuple[object, ...]], dict[str, object]]:
    rows = [_row_values(record) for record in records]
    _require(
        all(value is None or isinstance(value, str) for row in rows for value in row),
        "direct raw twin contains a non-string/non-blank cell",
    )
    kinds = Counter(str(record["kind"]) for record in records)
    blank_county = [record for record in records if record["county"] is None]
    pointers = Counter(record["distance"] for record in records)
    ordered_sha256 = _ordered_digest(rows)
    _require(len(rows) == 69_804, "raw record census drift")
    _require(kinds == Counter({"data": 68_806, "equate": 998}), "raw kind census drift")
    _require(
        len(blank_county) == 46
        and all(record["kind"] == "equate" for record in blank_county),
        "blank-County records are not exactly 46 EQUATES annotations",
    )
    _require(
        pointers["*P*"] == 283 and pointers["-------->"] == 282,
        "raw pointer-token census drift",
    )
    _require(
        ordered_sha256 == EXPECTED_ORDERED_ROWS_SHA256,
        f"direct raw twin row order/content drift: {ordered_sha256}",
    )
    return rows, {
        "raw_records": len(rows),
        "data_records": kinds["data"],
        "equate_records": kinds["equate"],
        "blank_county_equates": len(blank_county),
        "projectable_records": len(rows) - len(blank_county),
        "pointer_P": pointers["*P*"],
        "pointer_arrow": pointers["-------->"],
        "pointer_total": pointers["*P*"] + pointers["-------->"],
        "ordered_rows_sha256": ordered_sha256,
        "route_counts": dict(sorted(Counter(record["route"] for record in records).items())),
    }


def _text_cell(sheet, value: str, *, header: bool = False) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    cell.data_type = "s"
    if header:
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def _unpack_xlsx(payload: bytes, label: str) -> dict[str, bytes]:
    with ZipFile(io.BytesIO(payload), "r") as package:
        names = package.namelist()
        _require(len(names) == len(set(names)), f"{label} XLSX has duplicate package members")
        members = {name: package.read(name) for name in names}
        _require(package.testzip() is None, f"{label} XLSX CRC failure")
    _require("docProps/core.xml" in members, f"{label} XLSX has no core properties")
    return members


def _core_tree(payload: bytes) -> tuple[ElementTree.Element, dict[str, ElementTree.Element]]:
    try:
        root = ElementTree.fromstring(payload)
    except ElementTree.ParseError as error:
        raise DirectTwinError(f"invalid XLSX core-properties XML: {error}") from error
    nodes: dict[str, ElementTree.Element] = {}
    for label in ("created", "modified"):
        matched = root.findall(f".//{{{CORE_NAMESPACES['dcterms']}}}{label}")
        _require(len(matched) == 1, f"core-properties {label} node census: {len(matched)}")
        _require(
            matched[0].get(f"{{{CORE_NAMESPACES['xsi']}}}type") == "dcterms:W3CDTF",
            f"core-properties {label} xsi:type drift",
        )
        nodes[label] = matched[0]
    return root, nodes


def _serialize_core(root: ElementTree.Element) -> bytes:
    for prefix, namespace in CORE_NAMESPACES.items():
        ElementTree.register_namespace(prefix, namespace)
    return ElementTree.tostring(
        root, encoding="utf-8", xml_declaration=True, short_empty_elements=True,
    )


def _rewrite_core_times(payload: bytes, *, created: str, modified: str) -> bytes:
    root, nodes = _core_tree(payload)
    nodes["created"].text = created
    nodes["modified"].text = modified
    rewritten = _serialize_core(root)
    _root, verified = _core_tree(rewritten)
    _require(
        verified["created"].text == created and verified["modified"].text == modified,
        "core-properties timestamp rewrite failed exact reopen",
    )
    return rewritten


def _core_time_contract(payload: bytes) -> dict[str, object]:
    _root, nodes = _core_tree(payload)
    return {
        "created_nodes": 1,
        "modified_nodes": 1,
        "created": nodes["created"].text,
        "modified": nodes["modified"].text,
        "exact_fixed": (
            nodes["created"].text == FIXED_CORE_TIME
            and nodes["modified"].text == FIXED_CORE_TIME
        ),
    }


def _pack_canonical_members(members: Mapping[str, bytes]) -> bytes:
    output = io.BytesIO()
    with ZipFile(
        output, "w", compression=ZIP_DEFLATED, compresslevel=9,
        strict_timestamps=True,
    ) as package:
        package.comment = b""
        for name in sorted(members):
            info = ZipInfo(name, date_time=FIXED_ZIP_TIMESTAMP)
            info.compress_type = ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            info.internal_attr = 0
            info.extra = b""
            info.comment = b""
            package.writestr(info, members[name], compress_type=ZIP_DEFLATED, compresslevel=9)
    return output.getvalue()


def _canonicalize_xlsx(raw_payload: bytes) -> tuple[bytes, dict[str, object]]:
    members = _unpack_xlsx(raw_payload, "raw")
    members["docProps/core.xml"] = _rewrite_core_times(
        members["docProps/core.xml"],
        created=FIXED_CORE_TIME, modified=FIXED_CORE_TIME,
    )
    core_contract = _core_time_contract(members["docProps/core.xml"])
    _require(core_contract["exact_fixed"] is True, "canonical core timestamps drift")
    canonical = _pack_canonical_members(members)
    with ZipFile(io.BytesIO(canonical), "r") as package:
        canonical_names = package.namelist()
        fixed_timestamps = all(info.date_time == FIXED_ZIP_TIMESTAMP for info in package.infolist())
        empty_metadata = all(not info.extra and not info.comment for info in package.infolist())
        _require(package.testzip() is None, "canonical XLSX CRC failure")
    _require(canonical_names == sorted(canonical_names), "canonical XLSX member order drift")
    _require(fixed_timestamps and empty_metadata, "canonical XLSX ZIP metadata drift")
    member_ledger = [{
        "name": name, "bytes": len(members[name]), "sha256": _sha_bytes(members[name]),
    } for name in sorted(members)]
    return canonical, {
        "package_members": len(member_ledger),
        "package_member_names_sorted": True,
        "fixed_zip_timestamps": fixed_timestamps,
        "empty_zip_member_extra_and_comments": empty_metadata,
        "core_timestamp_contract": core_contract,
        "member_ledger_sha256": _sha_bytes(_json_bytes(member_ledger)),
        "members": member_ledger,
    }


def _core_time_mutation_probe(canonical_payload: bytes) -> dict[str, object]:
    members = _unpack_xlsx(canonical_payload, "canonical baseline")
    expected = _core_time_contract(members["docProps/core.xml"])
    _require(expected["exact_fixed"] is True, "mutation baseline core contract drift")
    planted_time = "2001-02-03T04:05:06Z"
    members["docProps/core.xml"] = _rewrite_core_times(
        members["docProps/core.xml"],
        created=FIXED_CORE_TIME, modified=planted_time,
    )
    mutated_payload = _pack_canonical_members(members)
    mutated_contract = _core_time_contract(members["docProps/core.xml"])
    rejected = mutated_contract != expected and mutated_contract["exact_fixed"] is False
    _require(rejected, "planted core modified-time mutation escaped exact contract")
    recovered_payload, recovered_package = _canonicalize_xlsx(mutated_payload)
    _require(
        recovered_payload == canonical_payload,
        "canonicalization did not exactly recover the planted core-time mutation",
    )
    return {
        "label": "volatile docProps/core.xml modified timestamp",
        "planted_modified": planted_time,
        "expected_contract": expected,
        "mutated_contract": mutated_contract,
        "rejected_by_exact_contract": rejected,
        "recanonicalized_byte_identical": True,
        "mutated_artifact": _artifact_identity(mutated_payload),
        "recovered_artifact": _artifact_identity(recovered_payload),
        "recovered_member_ledger_sha256": recovered_package["member_ledger_sha256"],
    }


def _build_workbook(rows: Sequence[Sequence[object]]) -> tuple[bytes, dict[str, object]]:
    workbook = Workbook(write_only=True)
    workbook.properties.creator = "TSMIS Comparison Perfection Audit"
    workbook.properties.lastModifiedBy = "TSMIS Comparison Perfection Audit"
    workbook.properties.created = FIXED_DOCUMENT_TIME
    workbook.properties.modified = FIXED_DOCUMENT_TIME
    workbook.properties.title = "Highway Sequence authoritative raw TSN audit twin"
    workbook.properties.subject = "Direct-source audit input; not family acceptance"
    workbook.properties.description = (
        "All 69,804 authoritative raw TSN records with exhaustive external provenance."
    )
    sheet = workbook.create_sheet(SHEET_NAME)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{len(rows) + 1}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    widths = {"A": 10, "B": 10, "C": 12, "D": 10, "E": 6, "F": 6, "G": 22, "H": 72}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 22
    sheet.append([_text_cell(sheet, header, header=True) for header in HEADERS])
    for row in rows:
        sheet.append([
            None if value is None else _text_cell(sheet, str(value))
            for value in row
        ])
    raw = io.BytesIO()
    workbook.save(raw)
    workbook.close()
    canonical, package = _canonicalize_xlsx(raw.getvalue())
    return canonical, package


def _reopen_workbook(
    payload: bytes, expected_rows: Sequence[Sequence[object]],
) -> dict[str, object]:
    required_parts = {
        "[Content_Types].xml", "_rels/.rels", "docProps/core.xml",
        "xl/workbook.xml", "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml", "xl/styles.xml",
    }
    with ZipFile(io.BytesIO(payload), "r") as package:
        names = package.namelist()
        _require(required_parts <= set(names), "canonical XLSX required-part gap")
        _require(package.testzip() is None, "canonical XLSX package failed reopen")
        _require(len(names) == len(set(names)), "canonical XLSX duplicate member")

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    try:
        _require(workbook.sheetnames == [SHEET_NAME], "workbook sheet universe drift")
        sheet = workbook[SHEET_NAME]
        streamed: list[tuple[object, ...]] = []
        formulas = errors = blank_physical_rows = extra_physical_cells = 0
        padded_trailing_blanks = 0
        max_width = 0
        for physical_row, cells in enumerate(sheet.iter_rows(), 1):
            max_width = max(max_width, len(cells))
            extra_physical_cells += max(0, len(cells) - len(HEADERS))
            values = tuple(cell.value for cell in cells[:len(HEADERS)])
            formulas += sum(cell.data_type == "f" for cell in cells)
            errors += sum(cell.data_type == "e" for cell in cells)
            if not any(value is not None for value in values):
                blank_physical_rows += 1
            if len(values) < len(HEADERS):
                padded_trailing_blanks += len(HEADERS) - len(values)
                values += (None,) * (len(HEADERS) - len(values))
            streamed.append(values)
        _require(len(streamed) == 69_805, "streamed workbook physical-row drift")
        _require(streamed[0] == HEADERS, "streamed workbook header drift")
        reopened = streamed[1:]
        _require(len(reopened) == len(expected_rows), "streamed workbook data-row drift")
        _require(
            all(actual == tuple(expected) for actual, expected in zip(reopened, expected_rows, strict=True)),
            "streamed workbook typed cell drift",
        )
        reopened_digest = _ordered_digest(reopened)
        _require(reopened_digest == EXPECTED_ORDERED_ROWS_SHA256, "reopened row digest drift")
        _require(not formulas and not errors, "workbook contains formula/error cells")
        _require(not blank_physical_rows and not extra_physical_cells, "workbook physical shape drift")
        style_workbook = load_workbook(
            io.BytesIO(payload), read_only=False, data_only=False,
        )
        try:
            style_sheet = style_workbook[SHEET_NAME]
            header = next(style_sheet.iter_rows(
                min_row=1, max_row=1, min_col=1, max_col=8,
            ))
            header_style_exact = all(
                cell.font.bold and cell.font.color is not None
                and cell.fill.fill_type == "solid"
                for cell in header
            )
            render_properties = {
                "freeze_panes": style_sheet.freeze_panes,
                "auto_filter": style_sheet.auto_filter.ref,
                "description_column_width": style_sheet.column_dimensions["H"].width,
                "gridlines_hidden": style_sheet.sheet_view.showGridLines is False,
            }
        finally:
            style_workbook.close()
        _require(header_style_exact, "render-facing header style drift")
        _require(render_properties["freeze_panes"] == "A2", "workbook freeze-pane drift")
        _require(render_properties["auto_filter"] == "A1:H69805", "workbook filter extent drift")
        render_surface = [list(row) for row in streamed[:12]]
        return {
            "zip_test_passed": True,
            "required_parts_present": True,
            "duplicate_package_members": 0,
            "sheet_names_exact": True,
            "sheet_name": SHEET_NAME,
            "streamed_physical_rows": len(streamed),
            "streamed_data_rows": len(reopened),
            "streamed_logical_columns": len(HEADERS),
            "maximum_streamed_physical_width": max_width,
            "padded_omitted_trailing_blank_cells": padded_trailing_blanks,
            "extra_physical_cells": extra_physical_cells,
            "blank_physical_rows": blank_physical_rows,
            "formulas": formulas,
            "errors": errors,
            "typed_rows_exact": True,
            "ordered_rows_sha256": reopened_digest,
            "freeze_panes": render_properties["freeze_panes"],
            "auto_filter": render_properties["auto_filter"],
            "render_surface_probe": {
                "range": "A1:H12",
                "values_sha256": _sha_bytes(_json_bytes(render_surface)),
                "header_style_exact": header_style_exact,
                **render_properties,
            },
        }
    finally:
        workbook.close()


def _provenance_document(
    records: Sequence[Mapping[str, object]],
    rows: Sequence[Sequence[object]],
    documents: Sequence[Mapping[str, object]],
    source_ordinals: Mapping[int, Mapping[str, int]],
    source_capture: Mapping[str, object],
    topology: Mapping[str, object],
    stage6_chain: Mapping[str, object],
) -> dict[str, object]:
    provenance_rows = []
    for workbook_row, (record, values) in enumerate(
        zip(records, rows, strict=True), 2,
    ):
        provenance_rows.append({
            "workbook_row": workbook_row,
            "source_ordinals": source_ordinals[id(record)],
            "source_ref": {
                key: record[key] for key in (
                    "member", "physical_page", "printed_page", "line", "top",
                )
            },
            "source_context": {
                key: record[key] for key in (
                    "district", "route", "direction", "kind", "raw_text",
                )
            },
            "workbook_values": list(values),
        })
    return {
        "schema_version": 1,
        "audit": "Highway Sequence direct-source raw-TSN audit twin provenance",
        "artifact_status": "DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE",
        "acceptance_eligible": False,
        "stage8_family_accepted": False,
        "reason": (
            "This exhaustive direct-source twin is an input to later product and "
            "evidence certification; it is not a detached family acceptance."
        ),
        "schema": {"sheet": SHEET_NAME, "headers": list(HEADERS)},
        "source_capture": source_capture,
        "accepted_stage6_chain": stage6_chain,
        "raw_documents": documents,
        "bidirectional_equate_topology": topology,
        "row_count": len(provenance_rows),
        "rows": provenance_rows,
    }


def _write_new_atomic(path: Path, payload: bytes) -> None:
    _require(not path.exists(), f"refusing to overwrite output artifact: {path}")
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent,
    )
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        _require(not path.exists(), f"output appeared during atomic write: {path}")
        os.replace(temporary_name, path)
    except BaseException:
        try:
            os.unlink(temporary_name)
        except FileNotFoundError:
            pass
        raise


def run(output_root: Path) -> dict[str, object]:
    output_root, output_gate = _validate_output_root(output_root)
    _require(
        importlib.metadata.version("openpyxl") == "3.1.5"
        and importlib.metadata.version("pdfplumber") == "0.11.9",
        "audit package-version drift",
    )
    _require(
        not any(
            name == "compare_highway_sequence_tsn"
            or name.endswith(".compare_highway_sequence_tsn")
            or name == "consolidate_tsn_highway_sequence"
            or name.endswith(".consolidate_tsn_highway_sequence")
            for name in sys.modules
        ),
        "product Highway Sequence code loaded before direct-source audit",
    )

    self_before = _capture_unbound_once(Path(__file__))
    # The source PDFs and accepted chain are each captured exactly once below.
    static_identities, static_payloads = _capture_static_inputs()
    captured, non_source, source_capture = _capture_source_universe()
    accepted_stage6, stage6_chain = _validate_stage6_chain(
        static_identities, static_payloads, source_capture,
    )

    records, documents, document_records, source_ordinals = _parse_captured_sources(captured)
    rows, counts = _validate_rows(records)
    fresh_stage6 = _fresh_stage6_digests(records, documents, accepted_stage6)

    topology = _event_topology(document_records)
    _require(
        topology["equate_rows"] == 998
        and topology["data_e_rows"] == 998
        and topology["forward_paired"] == 998
        and topology["reverse_paired"] == 998
        and topology["forward_unpaired"] == 0
        and topology["reverse_unpaired"] == 0
        and topology["pair_ledgers_exact"] is True,
        "bidirectional raw EQUATES/data-E topology drift",
    )
    reverse_mutation = _reverse_topology_mutation(document_records, topology)

    for member in [*captured, non_source]:
        _require(
            _stat_signature(Path(member.canonical_path)) == member.stat_signature,
            f"captured source changed after parsing: {member.name}",
        )

    workbook_payload, package = _build_workbook(rows)
    time.sleep(1.2)
    delayed_workbook_payload, delayed_package = _build_workbook(rows)
    _require(
        delayed_workbook_payload == workbook_payload and delayed_package == package,
        "delayed in-memory workbook rebuild is not byte-identical",
    )
    delayed_determinism = {
        "delay_milliseconds": 1_200,
        "first": _artifact_identity(workbook_payload),
        "second": _artifact_identity(delayed_workbook_payload),
        "package_contract_exact": delayed_package == package,
        "byte_identical": True,
    }
    core_time_mutation = _core_time_mutation_probe(workbook_payload)
    workbook_reopen = _reopen_workbook(workbook_payload, rows)
    provenance = _provenance_document(
        records, rows, documents, source_ordinals,
        source_capture, topology, stage6_chain,
    )
    provenance_payload = _json_bytes(provenance)

    self_after_payload = Path(__file__).read_bytes()
    _require(
        len(self_after_payload) == self_before.bytes
        and _sha_bytes(self_after_payload) == self_before.sha256,
        "direct-source builder changed during execution",
    )
    runtime = {
        "python": sys.version.split()[0],
        "openpyxl": importlib.metadata.version("openpyxl"),
        "pdfplumber": importlib.metadata.version("pdfplumber"),
    }
    artifacts = {
        "workbook": _artifact_identity(workbook_payload),
        "provenance": _artifact_identity(provenance_payload),
    }
    manifest = {
        "schema_version": 1,
        "audit": "Highway Sequence direct-source raw-TSN audit twin manifest",
        "artifact_status": "DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE",
        "acceptance_eligible": False,
        "stage8_family_accepted": False,
        "output_root_embedded": False,
        "output_artifact_names": list(OUTPUT_NAMES),
        "output_root_gate": output_gate,
        "independence": {
            "development_row_cache_read": False,
            "development_twin_read": False,
            "product_code_imported": False,
            "source_parser": "exact-bound accepted audit-owned Stage-6 parser",
        },
        "builder_identity": self_before.public_identity(),
        "runtime": runtime,
        "static_bindings": static_identities,
        "source_capture": source_capture,
        "accepted_stage6_chain": stage6_chain,
        "fresh_stage6_digests": fresh_stage6,
        "counts": counts,
        "bidirectional_equate_topology": topology,
        "reverse_topology_mutation": reverse_mutation,
        "workbook_schema": {"sheet": SHEET_NAME, "headers": list(HEADERS)},
        "canonical_xlsx_package": package,
        "delayed_in_memory_determinism": delayed_determinism,
        "core_timestamp_mutation": core_time_mutation,
        "workbook_reopen": workbook_reopen,
        "generated": artifacts,
    }
    manifest_payload = _json_bytes(manifest)
    artifacts_with_manifest = {
        **artifacts, "manifest": _artifact_identity(manifest_payload),
    }
    result = {
        "schema_version": 1,
        "audit": "Highway Sequence direct-source raw-TSN audit twin builder",
        "status": "PASS_DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE",
        "terminal": True,
        "artifact_status": "DIRECT_SOURCE_AUDIT_INPUT_NOT_FAMILY_ACCEPTANCE",
        "acceptance_eligible": False,
        "stage8_family_accepted": False,
        "reason": (
            "Direct authoritative-source fixture is complete, but product, "
            "Comparison/evidence, permanent-gate, detached-decision, and replay "
            "layers remain external."
        ),
        "output_root_embedded": False,
        "artifact_universe": list(OUTPUT_NAMES),
        "artifacts": artifacts_with_manifest,
        "counts": counts,
        "bidirectional_equate_topology": {
            key: topology[key] for key in (
                "equate_rows", "data_e_rows", "forward_paired", "reverse_paired",
                "forward_unpaired", "reverse_unpaired", "forward_ledger_sha256",
                "reverse_ledger_sha256", "pair_ledgers_exact",
            )
        },
        "reverse_topology_mutation_rejected": reverse_mutation["rejected"],
        "ordered_rows_sha256": counts["ordered_rows_sha256"],
        "invariants": {
            "source_file_universe_exact": True,
            "twelve_source_payloads_single_captured_and_exact_bound": True,
            "parser_consumed_exact_captured_payloads": True,
            "accepted_stage6_chain_exact": True,
            "fresh_stage6_digests_exact": fresh_stage6["all_equal_accepted_stage6"],
            "raw_records_69804": counts["raw_records"] == 69_804,
            "data_records_68806": counts["data_records"] == 68_806,
            "equate_records_998": counts["equate_records"] == 998,
            "blank_county_equates_46": counts["blank_county_equates"] == 46,
            "raw_pointer_tokens_565": counts["pointer_total"] == 565,
            "forward_equates_to_e_exact_998": topology["forward_paired"] == 998,
            "reverse_e_to_equates_exact_998": topology["reverse_paired"] == 998,
            "zero_unpaired_both_directions": (
                topology["forward_unpaired"] == topology["reverse_unpaired"] == 0
            ),
            "reverse_only_mutation_rejected": reverse_mutation["rejected"],
            "canonical_xlsx_bytes": package["fixed_zip_timestamps"],
            "core_created_modified_exact_fixed": package[
                "core_timestamp_contract"
            ]["exact_fixed"],
            "delayed_in_memory_xlsx_byte_identical": delayed_determinism[
                "byte_identical"
            ],
            "planted_core_time_mutation_rejected": core_time_mutation[
                "rejected_by_exact_contract"
            ],
            "planted_core_time_mutation_recanonicalized_exact": core_time_mutation[
                "recanonicalized_byte_identical"
            ],
            "xlsx_stream_reopen_exact": workbook_reopen["typed_rows_exact"],
            "xlsx_no_formulas_or_errors": not workbook_reopen["formulas"] and not workbook_reopen["errors"],
            "render_surface_probe_exact": workbook_reopen["render_surface_probe"]["header_style_exact"],
            "no_output_root_volatility": True,
            "output_root_disjointness_and_alias_controls": all(output_gate.values()),
            "not_family_acceptance": True,
        },
    }
    _require(all(result["invariants"].values()), "terminal direct-twin invariant failed")
    result_payload = _json_bytes(result)

    output_root.mkdir(parents=False, exist_ok=False)
    output_payloads = {
        WORKBOOK_NAME: workbook_payload,
        PROVENANCE_NAME: provenance_payload,
        MANIFEST_NAME: manifest_payload,
        RESULT_NAME: result_payload,
    }
    for name in OUTPUT_NAMES:
        _write_new_atomic(output_root / name, output_payloads[name])
    observed_names = sorted(path.name for path in output_root.iterdir())
    _require(observed_names == sorted(OUTPUT_NAMES), "terminal output artifact universe drift")
    for name, payload in output_payloads.items():
        _require((output_root / name).read_bytes() == payload, f"terminal reopen drift: {name}")

    print(json.dumps({
        "status": result["status"],
        "output_root": str(output_root),
        "artifacts": {
            name: _artifact_identity(payload) for name, payload in output_payloads.items()
        },
        "rows": counts["raw_records"],
        "forward_pairs": topology["forward_paired"],
        "reverse_pairs": topology["reverse_paired"],
    }, sort_keys=True))
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-root", type=Path, required=True,
        help="Required nonexistent output directory; its parent must already exist.",
    )
    arguments = parser.parse_args()
    run(arguments.output_root)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (DirectTwinError, stage6.ConservationError, OSError, ValueError, TypeError, KeyError) as error:
        print(f"FAIL Highway Sequence direct-source raw-TSN twin: {error}")
        raise SystemExit(1)
