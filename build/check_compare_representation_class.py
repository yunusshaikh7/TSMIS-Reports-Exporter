"""HF-09 / PCOA-FINAL-013: the representation-only class must be DISCLOSED and
nothing else.

Independent classification of the frozen corpus found cells that are the same
text once punctuation, spacing, quoting and letter case are set aside -
"NEVADA STATE LINE , END OF COUNTY" vs "NEVADA STATE LINE /END OF COUNTY"
(Highway Log, 1,243 per format), "SLO SB CO LINE" vs "SLO/SB CO LINE" and
"CITRUS AVE OC 54-1293" vs "Citrus Ave OC 54-1293" (Highway Sequence),
"NB OFF TO S. GEYSERVILLE" vs "NB OFF TO S.GEYSERVILLE" (Ramp Detail),
"''F'' ST" vs '"F" ST' (Intersection Detail - the KER 046 pair the product
already annotates through the evidence _quote_note), and a landmark's leading
apostrophe (Clean Road).

They are REAL literal differences between two sources, and the owner ruled on
2026-07-26 that they stay counted and flagged. The defect is only that an
unqualified headline total does not say how much of itself is presentation.
So this bundle adds a COUNT LINE, and this check exists to prove that is all
it added:

  * a representation-only cell is still counted, still `D`, still in every
    published total - and is now also reported separately;
  * a substantive change in the same column is never folded into that class;
  * turning the label off reproduces byte-for-byte the same cells, state
    masks, counts and typed outcome, so no verdict can depend on it;
  * the five audited vs-TSN families declare the column, and the TSMIS-vs-TSMIS
    flavors that inherit their schemas do NOT (the opposite ruling governs
    there - see check_compare_highway_sequence_equate.py).

CI-safe: pure Python fixtures, no local data.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_representation_class.py
"""
import shutil
import sys
import tempfile
from dataclasses import replace
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(ROOT / "scripts"), str(ROOT)]

from openpyxl import load_workbook                       # noqa: E402

import compare_clean_highway_tsn as ch                   # noqa: E402
import compare_env as ce                                 # noqa: E402
import compare_highway_log as hl                         # noqa: E402
import compare_highway_log_pdf as hlp                    # noqa: E402
import compare_highway_sequence_pdf as hslp              # noqa: E402
import compare_highway_sequence_tsn as hs                # noqa: E402
import compare_intersection_detail_pdf as idp            # noqa: E402
import compare_intersection_detail_tsn as idt            # noqa: E402
import compare_ramp_detail_pdf as rdp                    # noqa: E402
import compare_ramp_detail_tsn as rd                     # noqa: E402
from compare_core import CompareSchema, representation_only, run_compare  # noqa: E402
from events import Events                                # noqa: E402

failures = []


def check(label, condition, detail=""):
    print(("OK   " if condition else "FAIL ") + label
          + (f" - {detail}" if detail and not condition else ""))
    if not condition:
        failures.append(label)


# --------------------------------------------------------------------------- #
# 1. the measured classes, and what must NOT join them
# --------------------------------------------------------------------------- #
print("The measured class (each pair was found on the frozen corpus):")
MEASURED = (
    ("Highway Log separator", "NEVADA STATE LINE , END OF COUNTY",
     "NEVADA STATE LINE /END OF COUNTY"),
    ("Highway Sequence separator", "SLO SB CO LINE", "SLO/SB CO LINE"),
    ("Highway Sequence letter case", "CITRUS AVE OC 54-1293",
     "Citrus Ave OC 54-1293"),
    ("Ramp Detail spacing", "NB OFF TO S. GEYSERVILLE",
     "NB OFF TO S.GEYSERVILLE"),
    ("Intersection Detail quoting", "''F'' ST", '"F" ST'),
    ("Clean Road landmark apostrophe", "'-VIA BIG BEAR BLVD-",
     "-VIA BIG BEAR BLVD-"),
)
for label, left, right in MEASURED:
    check(f"  {label} is the representation class",
          representation_only(left, right), f"{left!r} vs {right!r}")

print("Substantive differences must NEVER join it:")
SUBSTANTIVE = (
    ("a different landmark", "END R REALIGNMENT", "END X REALIGNMENT"),
    ("a different number", "BR 55-239", "BR 55-240"),
    ("a dropped word", "NB OFF TO DOHENY PK RD", "NB OFF TO PK RD"),
    ("blank against punctuation", "", "---"),
    ("punctuation against blank", "/", ""),
    ("a different route claim", "103 SEP 53-145", "1/103 SEP 53-145"),
)
for label, left, right in SUBSTANTIVE:
    check(f"  {label} is NOT the representation class",
          not representation_only(left, right), f"{left!r} vs {right!r}")
check("  two equal cells are never in the class (it labels DIFFERENCES)",
      not representation_only("A B", "A B"))


# --------------------------------------------------------------------------- #
# 2. through the shipped engine: counted, disclosed, and never a verdict
# --------------------------------------------------------------------------- #
HEADER = ["Location", "Description", "Note"]
SCHEMA = CompareSchema(report_name="Fixture", header=HEADER,
                       side_a="TSMIS", side_b="TSN",
                       representation_fields=("Description",))
ROWS_A = [["001", "1.000", "NEVADA STATE LINE , END OF COUNTY", "x"],
          ["001", "2.000", "END R REALIGNMENT", "x"],
          ["001", "3.000", "SAME", "A , B"],
          ["001", "4.000", "SAME", "x"]]
ROWS_B = [["001", "1.000", "NEVADA STATE LINE /END OF COUNTY", "x"],
          ["001", "2.000", "END X REALIGNMENT", "x"],
          ["001", "3.000", "SAME", "A / B"],
          ["001", "4.000", "SAME", "x"]]


def run(schema, out):
    return run_compare(schema, ROWS_A, ROWS_B, True, str(out),
                       events=Events(), confirm_overwrite=lambda _p: True,
                       mode="values")


def summary_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple("" if v is None else str(v) for v in row)
                for row in wb["Summary"].iter_rows(values_only=True)]
    finally:
        wb.close()


def disclosure(path):
    for row in summary_rows(path):
        for i, value in enumerate(row):
            if value.startswith("— of which representation-only"):
                return row[i + 1]
    return None


tmp = Path(tempfile.mkdtemp(prefix="tsmis_repr_class_"))
try:
    print("Through run_compare with the label ON:")
    on = tmp / "on.xlsx"
    res_on = run(SCHEMA, on)
    counts_on = res_on.comparison_outcome.counts
    check("both Description differences are COUNTED (nothing is suppressed)",
          counts_on.differing_cells == 3 and counts_on.differing_rows == 3,
          f"{counts_on.differing_cells} cells / {counts_on.differing_rows} rows")
    check("the Summary discloses the representation-only subset, and only it",
          disclosure(on) == "1", str(disclosure(on)))
    check("the substantive change in the SAME column is not folded into it",
          disclosure(on) == "1")
    check("a representation-only difference in a column that did NOT opt in "
          "is not classified",
          disclosure(on) == "1")
    check("the Summary notes explain the class without claiming it is "
          "suppressed",
          any("Representation-only differences" in " ".join(row)
              and "COUNTED" in " ".join(row) for row in summary_rows(on)))

    print("With the label OFF (the byte-for-byte control):")
    off = tmp / "off.xlsx"
    res_off = run(replace(SCHEMA, representation_fields=()), off)
    counts_off = res_off.comparison_outcome.counts
    check("counts are identical",
          (counts_on.differing_cells, counts_on.differing_rows,
           counts_on.paired_rows, counts_on.side_a_only_rows,
           counts_on.side_b_only_rows, counts_on.per_field_counts)
          == (counts_off.differing_cells, counts_off.differing_rows,
              counts_off.paired_rows, counts_off.side_a_only_rows,
              counts_off.side_b_only_rows, counts_off.per_field_counts))
    check("the typed verdict is identical",
          (res_on.verdict, res_on.completion, res_on.status)
          == (res_off.verdict, res_off.completion, res_off.status))
    check("the label adds NOTHING to the disclosure-free workbook",
          disclosure(off) is None, str(disclosure(off)))

    def sheet(path, name):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return [tuple(row) for row in wb[name].iter_rows(values_only=True)]
        finally:
            wb.close()

    for name in ("Comparison", "TSMIS", "TSN", "Routes"):
        check(f"the published '{name}' sheet is cell-for-cell identical",
              sheet(on, name) == sheet(off, name))
    # The hidden state masks own discrepancy truth - they must not move either.
    check("the very-hidden E2 snapshot sheets are identical",
          sheet(on, "__CMP_E2_SNAPSHOT_A") == sheet(off, "__CMP_E2_SNAPSHOT_A")
          and sheet(on, "__CMP_E2_SNAPSHOT_B")
          == sheet(off, "__CMP_E2_SNAPSHOT_B"))
finally:
    shutil.rmtree(tmp, ignore_errors=True)


# --------------------------------------------------------------------------- #
# 3. the wiring: exactly the audited scope, on both editions
# --------------------------------------------------------------------------- #
print("Wiring - the five audited vs-TSN families declare their column:")
IN_SCOPE = (
    ("Highway Log vs TSN", hl._SCHEMA, ("Description",)),
    ("Highway Log (PDF) vs TSN", hlp.TSMIS_PDF_VS_TSN._schema, ("Description",)),
    ("Highway Sequence vs TSN", hs._SCHEMA, ("Description",)),
    ("Highway Sequence (PDF) vs TSN", hslp.TSMIS_PDF_VS_TSN._schema,
     ("Description",)),
    ("Intersection Detail vs TSN", idt._SCHEMA, ("Description",)),
    ("Intersection Detail (PDF) vs TSN", idp.TSMIS_PDF_VS_TSN._schema,
     ("Description",)),
    ("Ramp Detail vs TSN", rd._SCHEMA, ("Description",)),
    ("Ramp Detail (PDF) vs TSN", rdp.TSMIS_PDF_VS_TSN._schema, ("Description",)),
    ("Clean Road Highway", ch._SCHEMA, ("THY_LANDMARK_SHORT_DESC",)),
)
for label, schema, expected in IN_SCOPE:
    check(f"  {label} declares {expected[0]}",
          schema.representation_fields == expected,
          str(schema.representation_fields))
    check(f"  {label}'s column exists in its own header",
          all(name in schema.header for name in schema.representation_fields))

print("Wiring - TSMIS-vs-TSMIS flavors must NOT (the opposite ruling):")
OUT_OF_SCOPE = (
    ("Highway Log PDF vs Excel", hlp.TSMIS_PDF_VS_EXCEL._schema),
    ("Highway Sequence PDF vs Excel", hslp.TSMIS_PDF_VS_EXCEL._schema),
    ("Intersection Detail PDF vs Excel", idp.TSMIS_PDF_VS_EXCEL._schema),
    ("Ramp Detail PDF vs Excel", rdp.TSMIS_PDF_VS_EXCEL._schema),
    ("Highway Log cross-environment", ce._HL_BASE),
)
for label, schema in OUT_OF_SCOPE:
    check(f"  {label} does not", schema.representation_fields == (),
          str(schema.representation_fields))

if failures:
    print(f"\nFAILED {len(failures)} check(s):")
    for item in failures:
        print("  -", item)
    sys.exit(1)
print("\nOK  COMPARE-REPRESENTATION-CLASS: the measured punctuation / spacing "
      "/ quoting / letter-case class is classified and disclosed as a SUBSET "
      "of the published totals; every such cell stays counted and flagged, a "
      "substantive change never joins it, turning the label off reproduces "
      "every published cell, state mask, count and typed outcome exactly, and "
      "only the nine audited vs-TSN comparisons opt in.")
