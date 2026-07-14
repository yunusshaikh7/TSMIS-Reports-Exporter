#!/usr/bin/env python3
"""Classify Highway Sequence product-vs-source residuals for four TSN legs.

This is deliberately a NON-ACCEPTANCE development artifact.  It binds the
already frozen product publications and the cache-backed independent draft,
then independently reconstructs both pairing policies:

* source policy: complete PM identity; all five source fields, edit distance,
  then position for duplicate assignment;
* product policy: complete PM identity; asserted FT/Description mismatch count
  only, followed by the product's lexicographic exact assignment.

Every product count is rebuilt from the persisted Comparison ledger.  The
script does not import product comparison code and never treats a returned
aggregate count as evidence.  Final family acceptance must reparse immutable
raw Excel/PDF/TSN sources rather than promote this cache-backed classifier.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
from functools import lru_cache
import hashlib
from io import BytesIO
import json
import os
from pathlib import Path
import re
import stat
import tempfile
from typing import Callable, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils.escape import unescape as xlsx_unescape


VISUAL_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd"
)
DEFAULT_OUTPUT = (
    VISUAL_ROOT / "phase8_highway_sequence_product_residuals_draft_r1.json"
)

SOURCE_ROWS = VISUAL_ROOT / "phase8_highway_sequence_source_rows_draft_r1.json"
TSN_ROWS = VISUAL_ROOT / "phase8_highway_sequence_tsn_rows_draft_r1.json"
SOURCE_ORACLE = VISUAL_ROOT / "phase8_highway_sequence_source_oracle_draft_r1.json"
COMPARISON_DRAFT = VISUAL_ROOT / "phase8_highway_sequence_comparison_draft_r1.json"
DESCRIPTION_PROBE = (
    VISUAL_ROOT / "phase8_highway_sequence_description_normalization_probe_r1.json"
)
EXCEL_ESCAPE_PROBE = (
    VISUAL_ROOT / "phase8_highway_sequence_installed_excel_escape_probe_r1.json"
)
NORMALIZED_PARITY = (
    VISUAL_ROOT / "phase8_highway_sequence_product_comparison_parity_r1.json"
)


def _leg_root(name: str) -> Path:
    return VISUAL_ROOT / name


LEG_SPECS: dict[str, dict[str, object]] = {
    "excel_vs_raw_tsn": {
        "left_dataset": "current_tsmis_excel",
        "right_dataset": "raw_tsn",
        "draft_leg": "excel_vs_raw_tsn_full_pm",
        "side_a": "TSMIS",
        "side_b": "TSN",
        "root": _leg_root(
            "phase8_highway_sequence_product_comparison_excel_vs_raw_tsn_dev_r1"
        ),
        "result_binding": {
            "bytes": 21_185,
            "sha256": "2691fe4a5d6d1ed757d788c16bed7226a7966db8c1950423daf194369e6ae58c",
        },
        "values_binding": {
            "bytes": 34_385_223,
            "sha256": "ed9212ad564d3b3ab600845cc0341913237958f948b0726de70f77bf7696e8cd",
        },
    },
    "excel_vs_normalized_tsn": {
        "left_dataset": "current_tsmis_excel",
        "right_dataset": "normalized_tsn",
        "draft_leg": "excel_vs_normalized_tsn_full_pm",
        "side_a": "TSMIS",
        "side_b": "TSN",
        "root": _leg_root(
            "phase8_highway_sequence_product_comparison_excel_vs_normalized_tsn_r2"
        ),
        "result_binding": {
            "bytes": 16_069,
            "sha256": "b1cf6f791c18917dfb51b3f9f2d8331075091992ce3d3c3415032108ee9bec83",
        },
        "values_binding": {
            "bytes": 34_366_704,
            "sha256": "bb2d7c911ad235649468d01e019bc5cf7c7d58c293957d6667915241caafc13b",
        },
    },
    "pdf_vs_raw_tsn": {
        "left_dataset": "current_tsmis_pdf",
        "right_dataset": "raw_tsn",
        "draft_leg": "pdf_vs_raw_tsn_full_pm",
        "side_a": "TSMIS (PDF)",
        "side_b": "TSN",
        "root": _leg_root(
            "phase8_highway_sequence_product_comparison_pdf_vs_raw_tsn_dev_r1"
        ),
        "result_binding": {
            "bytes": 21_346,
            "sha256": "31656c378240c30218054ae57972d5480f68aa37045140a2c9d6a3aa3e7b2b81",
        },
        "values_binding": {
            "bytes": 34_349_992,
            "sha256": "1d228e2ec418e19830262e383ac7a5968044271906cb5f80cf9daaf864396f69",
        },
    },
    "pdf_vs_normalized_tsn": {
        "left_dataset": "current_tsmis_pdf",
        "right_dataset": "normalized_tsn",
        "draft_leg": "pdf_vs_normalized_tsn_full_pm",
        "side_a": "TSMIS (PDF)",
        "side_b": "TSN",
        "root": _leg_root(
            "phase8_highway_sequence_product_comparison_pdf_vs_normalized_tsn_r2"
        ),
        "result_binding": {
            "bytes": 16_228,
            "sha256": "65d79577e9dbc7dfbce22d3d12fa4b8a670edb78b439b56b2802afeaa077a59a",
        },
        "values_binding": {
            "bytes": 34_331_218,
            "sha256": "23e9102461f2458866ef557efd4576411063f2522a5bc1f202a629fde751f180",
        },
    },
}


INPUT_BINDINGS: dict[str, tuple[Path, int, str]] = {
    "source_rows": (
        SOURCE_ROWS,
        49_304_637,
        "564cf21972aeaf461811095997524c2d02f3ca4f238bb8da8b715415df2762f8",
    ),
    "tsn_rows": (
        TSN_ROWS,
        28_829_216,
        "b18d2e077b79920cb1f687f06f8193b25e1d8cd2ebeb1d071b84c22b372598a7",
    ),
    "source_oracle": (
        SOURCE_ORACLE,
        4_008_580,
        "2c0997b7d3eb000ac40eddcb5107fa86951ca98825b58394af2b640a5c964b90",
    ),
    "corrected_comparison_draft": (
        COMPARISON_DRAFT,
        113_580_300,
        "4198f7e4a65a4afbe164e738defaf36ec0270efc328f0e46d400937c7b9efb1c",
    ),
    "description_normalization_probe": (
        DESCRIPTION_PROBE,
        174_929,
        "202fcb82b6ba62d15fcd273b19f4f35de672d06da39fd710982ba65350e8bdd1",
    ),
    "installed_excel_escape_probe": (
        EXCEL_ESCAPE_PROBE,
        31_722,
        "ec8c61c8cb8e629abee82d83abaacc1b9c9ebc3ce4c2356a6da527e4ead42b07",
    ),
    "normalized_product_parity": (
        NORMALIZED_PARITY,
        42_381,
        "bb7c8550724b71e657781f86579e25b2f70c96bf8bf3380d049f70118f98961f",
    ),
}

VALUE_FIELDS = ("City", "HG", "FT", "Distance To Next Point", "Description")
ASSERTED_FIELDS = ("FT", "Description")
ASSERTED_INDEX = {"FT": 2, "Description": 4}
PRODUCT_STATE_FIELDS = (
    "County", "City", "HG", "FT", "Distance To Next Point", "Description"
)
PRODUCT_CONTEXT_FIELDS = frozenset(("City", "HG", "Distance To Next Point"))
PM_RE = re.compile(r"^([A-Z]?\d{3}\.\d{3})(E?)$")
ROUTE_PREFIX_RE = re.compile(r"^(\d{1,3}[A-Z]?)/")
OOXML_ESCAPE_RE = re.compile(r"_x[0-9a-f]{4}_", re.IGNORECASE)
LINK_RE = re.compile(
    r'^=HYPERLINK\("#(?P<sheet>(?:\'(?:[^\']|\'\')+\'|[^!]+))!'
    r'(?P<row>\d+):(?P=row)",(?P<label>\d+)\)$'
)
STATE_HEADER = "__CMP_E1_STATE_V1_C001_P0000_P0005"


class ResidualError(RuntimeError):
    """A frozen identity or a row-level reconciliation contract drifted."""


@dataclass(frozen=True)
class Row:
    dataset: str
    product_index: int
    source_ref: str
    route: str
    county: str
    pm_base: str
    pm_suffix: str
    values: tuple[str, ...]
    raw_description: object
    kind: str
    is_tsmis: bool
    is_excel: bool

    @property
    def pm_full(self) -> str:
        return self.pm_base + self.pm_suffix

    @property
    def key(self) -> tuple[str, str, str]:
        return self.route, self.county, self.pm_full

    @property
    def identity(self) -> list[str]:
        return [self.route, self.county, self.pm_full]

    @property
    def truth_values(self) -> tuple[str, ...]:
        values = list(self.values)
        values[-1] = _truth_description(self)
        return tuple(values)

    @property
    def product_values(self) -> tuple[str, ...]:
        values = list(self.values)
        values[-1] = _product_description(self.raw_description)
        return tuple(values)


@dataclass
class Pairing:
    pairs: set[tuple[int, int]]
    left_only: set[int]
    right_only: set[int]
    duplicate_traces: list[dict[str, object]]
    aggregate_cost: object


def _require(condition: object, message: str) -> None:
    if not condition:
        raise ResidualError(message)


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical(value: object, *, newline: bool = False) -> bytes:
    payload = json.dumps(
        value, ensure_ascii=False, allow_nan=False,
        sort_keys=True, separators=(",", ":"),
    ).encode("utf-8")
    return payload + (b"\n" if newline else b"")


def _ledger(value: object) -> dict[str, object]:
    payload = _canonical(value)
    return {"rows": len(value), "bytes": len(payload), "sha256": _sha_bytes(payload)}


def _absolute_unresolved(path: Path) -> Path:
    """Return an absolute spelling without dereferencing any path component."""
    return Path(os.path.abspath(os.fspath(path)))


def _guard_no_reparse(
    path: Path, label: str, *, allow_missing_leaf: bool = False,
) -> Path:
    """Reject symlinks/junctions before any operation can dereference them."""
    candidate = _absolute_unresolved(path)
    components = tuple(reversed((candidate, *candidate.parents)))
    reparse_mask = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    for component in components:
        try:
            metadata = os.lstat(component)
        except FileNotFoundError:
            if component == candidate and allow_missing_leaf:
                continue
            if component != candidate:
                continue
            raise ResidualError(f"{label}: path does not exist: {candidate}")
        is_reparse = bool(
            getattr(metadata, "st_file_attributes", 0) & reparse_mask
        )
        _require(
            not stat.S_ISLNK(metadata.st_mode) and not is_reparse,
            f"{label}: symlink/reparse path component is forbidden: {component}",
        )
    return candidate


def _guard_regular_file(path: Path, label: str) -> Path:
    candidate = _guard_no_reparse(path, label)
    try:
        metadata = os.lstat(candidate)
    except FileNotFoundError as exc:
        raise ResidualError(f"{label}: path does not exist: {candidate}") from exc
    _require(stat.S_ISREG(metadata.st_mode), f"{label}: not an ordinary file: {candidate}")
    return candidate


def _bound_artifact_paths() -> tuple[Path, ...]:
    paths = [binding[0] for binding in INPUT_BINDINGS.values()]
    for spec in LEG_SPECS.values():
        root = Path(spec["root"])
        paths.extend((
            root / "result.json",
            root / "comparison.xlsx",
            root / "comparison (values).xlsx",
        ))
    absolute = tuple(_absolute_unresolved(path) for path in paths)
    _require(len(set(map(os.path.normcase, map(str, absolute)))) == len(absolute),
             "bound artifact path universe contains a lexical alias")
    return absolute


def _guard_output_path(output: Path, protected: Sequence[Path]) -> Path:
    candidate = _guard_no_reparse(
        output, "classifier output", allow_missing_leaf=True,
    )
    if candidate.exists():
        _require(candidate.is_file(), f"classifier output is not a file: {candidate}")
    candidate_key = os.path.normcase(str(candidate))
    for protected_path in protected:
        protected_absolute = _absolute_unresolved(protected_path)
        _require(
            candidate_key != os.path.normcase(str(protected_absolute)),
            f"classifier output aliases a protected input: {protected_absolute}",
        )
        if candidate.exists() and protected_absolute.exists():
            try:
                same_file = os.path.samefile(candidate, protected_absolute)
            except OSError as exc:
                raise ResidualError(
                    "classifier output alias check failed: "
                    f"{type(exc).__name__}: {str(exc).splitlines()[0]}"
                ) from exc
            _require(
                not same_file,
                f"classifier output hardlink-aliases a protected input: {protected_absolute}",
            )
    return candidate


def _identity(path: Path) -> dict[str, object]:
    path = _guard_regular_file(path, f"identity input {path.name}")
    return {
        "path": str(path), "bytes": path.stat().st_size, "sha256": _sha_file(path),
    }


def _skip_record(exc: BaseException) -> dict[str, object]:
    record: dict[str, object] = {
        "status": "skipped",
        "exception_type": type(exc).__name__,
    }
    for name in ("errno", "winerror"):
        value = getattr(exc, name, None)
        if value is not None:
            record[name] = value
    return record


def _expect_residual_rejection(
    action: Callable[[], object], expected_fragment: str, label: str,
) -> None:
    try:
        action()
    except ResidualError as exc:
        _require(expected_fragment in str(exc), f"{label}: wrong rejection: {exc}")
    else:
        raise ResidualError(f"{label}: negative fixture was accepted")


def _path_and_alias_mutation_probes() -> dict[str, object]:
    """Exercise guards only inside a disposable directory; never touch witnesses."""
    with tempfile.TemporaryDirectory(prefix="hsl_residual_guard_") as temporary:
        root = Path(temporary)
        source = root / "source.bin"
        source.write_bytes(b"frozen-input")
        lexical_alias = source.parent / "." / source.name
        _expect_residual_rejection(
            lambda: _guard_output_path(lexical_alias, (source,)),
            "aliases a protected input", "lexical output alias",
        )

        hardlink_probe: dict[str, object]
        hardlink = root / "hardlink.bin"
        try:
            os.link(source, hardlink)
        except (OSError, NotImplementedError) as exc:
            hardlink_probe = _skip_record(exc)
        else:
            _expect_residual_rejection(
                lambda: _guard_output_path(hardlink, (source,)),
                "hardlink-aliases a protected input", "hardlink output alias",
            )
            hardlink_probe = {"status": "passed"}

        file_symlink_probe: dict[str, object]
        file_symlink = root / "file-symlink.bin"
        try:
            os.symlink(source, file_symlink)
        except (OSError, NotImplementedError) as exc:
            file_symlink_probe = _skip_record(exc)
        else:
            _expect_residual_rejection(
                lambda: _guard_regular_file(file_symlink, "file symlink fixture"),
                "symlink/reparse path component", "file symlink",
            )
            file_symlink_probe = {"status": "passed"}

        directory_symlink_probe: dict[str, object]
        target_directory = root / "target-directory"
        target_directory.mkdir()
        (target_directory / "inside.bin").write_bytes(b"inside")
        directory_symlink = root / "directory-symlink"
        try:
            os.symlink(target_directory, directory_symlink, target_is_directory=True)
        except (OSError, NotImplementedError) as exc:
            directory_symlink_probe = _skip_record(exc)
        else:
            _expect_residual_rejection(
                lambda: _guard_regular_file(
                    directory_symlink / "inside.bin", "directory symlink fixture",
                ),
                "symlink/reparse path component", "directory symlink component",
            )
            directory_symlink_probe = {"status": "passed"}

    return {
        "disposable_fixture_only": True,
        "lexical_output_alias": {"status": "passed"},
        "hardlink_output_alias": hardlink_probe,
        "file_symlink_or_reparse": file_symlink_probe,
        "directory_symlink_or_reparse_component": directory_symlink_probe,
        "unsupported_link_creation_is_truthfully_skipped": True,
    }


def _bind_inputs() -> dict[str, dict[str, object]]:
    bound: dict[str, dict[str, object]] = {}
    for label, (path, expected_bytes, expected_sha) in INPUT_BINDINGS.items():
        observed = _identity(path)
        _require(
            (observed["bytes"], observed["sha256"]) == (expected_bytes, expected_sha),
            f"{label}: frozen identity drift: {observed}",
        )
        bound[label] = observed
    for label, spec in LEG_SPECS.items():
        root = Path(spec["root"])
        result_path = root / "result.json"
        values_path = root / "comparison (values).xlsx"
        for suffix, path, expected in (
            ("result", result_path, spec["result_binding"]),
            ("values_workbook", values_path, spec["values_binding"]),
        ):
            observed = _identity(path)
            _require(
                (observed["bytes"], observed["sha256"])
                == (expected["bytes"], expected["sha256"]),
                f"{label} {suffix}: frozen identity drift: {observed}",
            )
            bound[f"{label}_{suffix}"] = observed
    return bound


def _strict_json(path: Path) -> dict[str, object]:
    raw = path.read_bytes()
    try:
        value = json.loads(raw.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise ResidualError(f"{path.name}: invalid UTF-8 JSON: {exc}") from exc
    _require(isinstance(value, dict), f"{path.name}: JSON root is not an object")
    return value


def _captured_json(
    path: Path, expected: Mapping[str, object], label: str,
) -> tuple[dict[str, object], dict[str, object]]:
    """Hash and parse one captured byte buffer; never hash/read different bytes."""
    path = _guard_regular_file(path, f"{label} JSON")
    raw = path.read_bytes()
    identity = {
        "path": str(path), "bytes": len(raw), "sha256": _sha_bytes(raw),
    }
    _require(
        (identity["bytes"], identity["sha256"])
        == (expected["bytes"], expected["sha256"]),
        f"{label}: captured JSON identity drift: {identity}",
    )
    try:
        value = json.loads(raw.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise ResidualError(f"{label}: invalid captured UTF-8 JSON: {exc}") from exc
    _require(isinstance(value, dict), f"{label}: captured JSON root is not an object")
    assert isinstance(value, dict)
    return value, identity


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _space(value: object, *, xlsx: bool = False) -> str:
    text = _text(value)
    if xlsx:
        text = xlsx_unescape(text)
    return " ".join(text.split())


def _route(value: object) -> str:
    match = re.fullmatch(r"(\d{1,3})([A-Za-z]?)", _text(value))
    _require(match is not None, f"invalid route: {value!r}")
    assert match is not None
    return match.group(1).zfill(3) + match.group(2).upper()


def _county(value: object) -> str:
    return _text(value).rstrip(".").upper()


def _pm(value: object) -> tuple[str, str]:
    text = _text(value).upper()
    if not text:
        return "", ""
    match = PM_RE.fullmatch(text)
    _require(match is not None, f"invalid complete postmile: {value!r}")
    assert match is not None
    return match.group(1), match.group(2)


def _truth_description(row: Row) -> str:
    text = _space(row.raw_description, xlsx=row.is_excel)
    if row.is_tsmis:
        match = ROUTE_PREFIX_RE.match(text)
        if match is not None and _route(match.group(1)) == row.route:
            text = text[match.end():].lstrip()
    return text


def _product_description(value: object) -> str:
    text = _space(value, xlsx=False)
    return ROUTE_PREFIX_RE.sub("", text, count=1).strip()


def _load_tsmis(
    serialized: Sequence[Mapping[str, object]], dataset: str,
) -> list[Row]:
    is_excel = "excel" in dataset
    rows: list[Row] = []
    for product_index, item in enumerate(serialized):
        raw = tuple(item["values"])
        route = _route(item["route"])
        pm_base = _text(raw[2]).upper() + _text(raw[3]).upper()
        _require(not pm_base or PM_RE.fullmatch(pm_base), f"bad TSMIS PM: {item}")
        suffix = _text(raw[4]).upper()
        _require(suffix in ("", "E"), f"bad TSMIS suffix: {item}")
        rows.append(Row(
            dataset=dataset,
            product_index=product_index,
            source_ref=str(item["source_ref"]),
            route=route,
            county=_county(raw[0]),
            pm_base=pm_base,
            pm_suffix=suffix,
            values=(
                _space(raw[1], xlsx=is_excel),
                _space(raw[5], xlsx=is_excel),
                _space(raw[6], xlsx=is_excel),
                _space(raw[7], xlsx=is_excel),
                _space(raw[8], xlsx=is_excel),
            ),
            raw_description=raw[8],
            kind="tsmis",
            is_tsmis=True,
            is_excel=is_excel,
        ))
    return rows


def _load_raw_tsn(
    serialized: Sequence[Mapping[str, object]],
) -> tuple[list[Row], list[Row], list[Row]]:
    all_rows: list[Row] = []
    known: list[Row] = []
    unknown: list[Row] = []
    for product_index, item in enumerate(serialized):
        pm_base, suffix = _pm(item["pm"])
        row = Row(
            dataset="raw_tsn",
            product_index=product_index,
            source_ref=(
                f"{item['member']}:page:{item['physical_page']}:line:{item['line']}"
            ),
            route=_route(item["route"]),
            county=_county(item.get("county")),
            pm_base=pm_base,
            pm_suffix=suffix,
            values=(
                _space(item.get("city")), _space(item.get("hg")),
                _space(item.get("ft")), _space(item.get("distance")),
                _space(item.get("description")),
            ),
            raw_description=item.get("description"),
            kind=str(item["kind"]),
            is_tsmis=False,
            is_excel=False,
        )
        all_rows.append(row)
        (known if row.county else unknown).append(row)
    return all_rows, known, unknown


def _load_normalized_tsn(
    serialized: Sequence[Mapping[str, object]],
) -> list[Row]:
    rows: list[Row] = []
    for product_index, item in enumerate(serialized):
        raw = tuple(item["values"])
        pm_base, suffix = _pm(raw[2])
        description = _space(raw[7])
        rows.append(Row(
            dataset="normalized_tsn",
            product_index=product_index,
            source_ref=f"normalized:row:{item['source_row']}",
            route=_route(raw[0]),
            county=_county(raw[1]),
            pm_base=pm_base,
            pm_suffix=suffix,
            values=(
                _space(raw[3]), _space(raw[4]), _space(raw[5]),
                _space(raw[6]), description,
            ),
            raw_description=raw[7],
            kind="equate" if description.startswith("EQUATES TO") else "data",
            is_tsmis=False,
            is_excel=False,
        ))
    return rows


@lru_cache(maxsize=None)
def _char_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for left_index, left_char in enumerate(left, 1):
        current = [left_index]
        for right_index, right_char in enumerate(right, 1):
            current.append(min(
                current[-1] + 1,
                previous[right_index] + 1,
                previous[right_index - 1] + (left_char != right_char),
            ))
        previous = current
    return previous[-1]


def _add_cost(left: object, right: object) -> object:
    if isinstance(left, tuple):
        assert isinstance(right, tuple)
        return tuple(a + b for a, b in zip(left, right, strict=True))
    return int(left) + int(right)


def _assignment(
    left: Sequence[Row], right: Sequence[Row],
    cost: Callable[[Row, Row, int, int], object],
    zero: object,
) -> tuple[list[tuple[Row, Row]], list[Row], list[Row], object]:
    if not left or not right:
        return [], list(left), list(right), zero
    swapped = len(left) > len(right)
    small = right if swapped else left
    large = left if swapped else right
    _require(len(large) <= 12, f"assignment group exceeds exact bound: {len(left)}x{len(right)}")

    @lru_cache(maxsize=None)
    def solve(index: int, used: int) -> tuple[object, tuple[int, ...]]:
        if index == len(small):
            return zero, ()
        best: tuple[object, tuple[int, ...]] | None = None
        for candidate in range(len(large)):
            if used & (1 << candidate):
                continue
            tail_cost, tail_vector = solve(index + 1, used | (1 << candidate))
            if swapped:
                pair_cost = cost(large[candidate], small[index], candidate, index)
            else:
                pair_cost = cost(small[index], large[candidate], index, candidate)
            value = (_add_cost(pair_cost, tail_cost), (candidate, *tail_vector))
            if best is None or value < best:
                best = value
        _require(best is not None, "assignment solver found no candidate")
        assert best is not None
        return best

    total, vector = solve(0, 0)
    pairs: list[tuple[Row, Row]] = []
    for small_index, large_index in enumerate(vector):
        if swapped:
            pairs.append((large[large_index], small[small_index]))
        else:
            pairs.append((small[small_index], large[large_index]))
    pairs.sort(key=lambda pair: (pair[0].product_index, pair[1].product_index))
    used_left = {id(pair[0]) for pair in pairs}
    used_right = {id(pair[1]) for pair in pairs}
    return (
        pairs,
        [row for row in left if id(row) not in used_left],
        [row for row in right if id(row) not in used_right],
        total,
    )


def _independent_cost(left: Row, right: Row, lp: int, rp: int) -> tuple[int, int, int]:
    left_values = left.truth_values
    right_values = right.truth_values
    return (
        sum(a != b for a, b in zip(left_values, right_values, strict=True)),
        sum(_char_distance(a, b) for a, b in zip(left_values, right_values, strict=True)),
        abs(lp - rp),
    )


def _product_cost(left: Row, right: Row, _lp: int, _rp: int) -> int:
    left_values = left.product_values
    right_values = right.product_values
    return sum(
        left_values[ASSERTED_INDEX[field]] != right_values[ASSERTED_INDEX[field]]
        for field in ASSERTED_FIELDS
    )


def _pair_rows(
    left: Sequence[Row], right: Sequence[Row], *, product: bool,
) -> Pairing:
    left_groups: dict[tuple[str, str, str], list[Row]] = defaultdict(list)
    right_groups: dict[tuple[str, str, str], list[Row]] = defaultdict(list)
    for row in left:
        left_groups[row.key].append(row)
    for row in right:
        right_groups[row.key].append(row)
    pairs: set[tuple[int, int]] = set()
    left_only: set[int] = set()
    right_only: set[int] = set()
    traces: list[dict[str, object]] = []
    cost_fn = _product_cost if product else _independent_cost
    zero: object = 0 if product else (0, 0, 0)
    aggregate_cost: object = zero
    for identity in sorted(set(left_groups) | set(right_groups)):
        group_left = left_groups.get(identity, [])
        group_right = right_groups.get(identity, [])
        group_pairs, group_left_only, group_right_only, total = _assignment(
            group_left, group_right, cost_fn, zero,
        )
        aggregate_cost = _add_cost(aggregate_cost, total)
        pairs.update((a.product_index, b.product_index) for a, b in group_pairs)
        left_only.update(row.product_index for row in group_left_only)
        right_only.update(row.product_index for row in group_right_only)
        if len(group_left) > 1 or len(group_right) > 1:
            traces.append({
                "identity": list(identity),
                "left": [row.source_ref for row in group_left],
                "right": [row.source_ref for row in group_right],
                "pairs": [[a.source_ref, b.source_ref] for a, b in group_pairs],
                "left_only": [row.source_ref for row in group_left_only],
                "right_only": [row.source_ref for row in group_right_only],
                "cost": list(total) if isinstance(total, tuple) else total,
            })
    return Pairing(pairs, left_only, right_only, traces, aggregate_cost)


def _row_by_index(rows: Sequence[Row]) -> dict[int, Row]:
    result = {row.product_index: row for row in rows}
    _require(len(result) == len(rows), "duplicate product index in source rows")
    return result


def _pair_sort_key(
    pair: tuple[int, int], left_by_index: Mapping[int, Row],
) -> tuple[object, ...]:
    return left_by_index[pair[0]].route, pair[0], pair[1]


def _diff_fields(
    left: Row, right: Row, *, product: bool,
) -> tuple[str, ...]:
    values_a = left.product_values if product else left.truth_values
    values_b = right.product_values if product else right.truth_values
    return tuple(
        field for field in ASSERTED_FIELDS
        if values_a[ASSERTED_INDEX[field]] != values_b[ASSERTED_INDEX[field]]
    )


def _all_diff_fields(left: Row, right: Row) -> tuple[str, ...]:
    return tuple(
        field for field, value_a, value_b
        in zip(VALUE_FIELDS, left.truth_values, right.truth_values, strict=True)
        if value_a != value_b
    )


def _counts_for_pairs(
    pairs: Iterable[tuple[int, int]], left_by_index: Mapping[int, Row],
    right_by_index: Mapping[int, Row], *, product: bool,
) -> dict[str, object]:
    field_counts: Counter[str] = Counter()
    differing_rows = 0
    pair_count = 0
    for left_index, right_index in pairs:
        pair_count += 1
        fields = _diff_fields(
            left_by_index[left_index], right_by_index[right_index], product=product,
        )
        field_counts.update(fields)
        differing_rows += bool(fields)
    return {
        "paired_rows": pair_count,
        "differing_rows": differing_rows,
        "differing_cells": sum(field_counts.values()),
        "per_field_counts": {
            field: field_counts[field] for field in ASSERTED_FIELDS
        },
    }


def _normalize_returned_counts(value: object, label: str) -> dict[str, object]:
    _require(isinstance(value, dict), f"{label}: absent returned counts")
    assert isinstance(value, dict)
    per_field = value.get("per_field_counts")
    _require(isinstance(per_field, dict), f"{label}: absent per-field counts")
    assert isinstance(per_field, dict)
    normalized_fields = Counter()
    for raw_field, raw_count in per_field.items():
        field = str(raw_field).split(":", 1)[-1]
        _require(type(raw_count) is int and raw_count >= 0, f"{label}: bad count")
        if raw_count:
            normalized_fields[field] += raw_count
    return {
        "known": value.get("known"),
        "paired_rows": value.get("paired_rows"),
        "side_a_only_rows": value.get("side_a_only_rows"),
        "side_b_only_rows": value.get("side_b_only_rows"),
        "differing_rows": value.get("differing_rows"),
        "differing_cells": value.get("differing_cells"),
        "per_field_counts": dict(sorted(normalized_fields.items())),
        "asserted_cells": value.get("asserted_cells"),
        "context_cells": value.get("context_cells"),
    }


def _link_index(cell, expected_sheet: str, label: str) -> int:
    _require(cell.data_type == "f" and isinstance(cell.value, str), f"{label}: bad link")
    match = LINK_RE.fullmatch(cell.value)
    _require(match is not None, f"{label}: malformed link {cell.value!r}")
    assert match is not None
    sheet = match.group("sheet")
    if sheet.startswith("'"):
        sheet = sheet[1:-1].replace("''", "'")
    row = int(match.group("row"))
    _require(
        sheet == expected_sheet and row == int(match.group("label")) and row >= 2,
        f"{label}: link target drift {cell.value!r}",
    )
    return row - 2


def _scan_product_workbook(
    path: Path, side_a: str, side_b: str,
    left: Sequence[Row], right: Sequence[Row],
    expected_binding: Mapping[str, object],
) -> dict[str, object]:
    left_by_index = _row_by_index(left)
    right_by_index = _row_by_index(right)
    resolved = _guard_regular_file(path, f"{path.name} workbook")
    raw_workbook = resolved.read_bytes()
    captured_identity = {
        "path": str(resolved), "bytes": len(raw_workbook),
        "sha256": _sha_bytes(raw_workbook),
    }
    _require(
        (captured_identity["bytes"], captured_identity["sha256"])
        == (expected_binding["bytes"], expected_binding["sha256"]),
        f"{path.name}: captured workbook identity drift",
    )
    stream = BytesIO(raw_workbook)
    workbook = load_workbook(stream, read_only=True, data_only=False)
    try:
        _require("Comparison" in workbook.sheetnames, f"{path.name}: no Comparison sheet")
        worksheet = workbook["Comparison"]
        physical = iter(worksheet.iter_rows())
        header = tuple(cell.value for cell in next(physical, ()))
        expected_header = (
            "Route", "PM", "#", f"{side_a} Row", f"{side_b} Row",
            "Status", "Diffs", *PRODUCT_STATE_FIELDS, STATE_HEADER,
        )
        _require(header == expected_header, f"{path.name}: Comparison header drift")

        pairs: set[tuple[int, int]] = set()
        left_only: set[int] = set()
        right_only: set[int] = set()
        pair_diffs: dict[tuple[int, int], tuple[str, ...]] = {}
        used_left: Counter[int] = Counter()
        used_right: Counter[int] = Counter()
        row_ids: list[list[object]] = []
        for physical_row, cells in enumerate(physical, 2):
            cells = tuple(cells)
            values = tuple(cell.value for cell in cells)
            if not any(value is not None for value in values):
                continue
            _require(len(cells) == len(expected_header), f"Comparison row {physical_row}: width")
            route = _text(values[0])
            visible_key = _text(values[1])
            occurrence = values[2]
            _require(type(occurrence) is int and occurrence >= 1, f"row {physical_row}: occurrence")
            index_a = (
                _link_index(cells[3], side_a, f"row {physical_row} side A")
                if values[3] is not None else None
            )
            index_b = (
                _link_index(cells[4], side_b, f"row {physical_row} side B")
                if values[4] is not None else None
            )
            _require(index_a is not None or index_b is not None, f"row {physical_row}: empty")
            left_row = left_by_index.get(index_a) if index_a is not None else None
            right_row = right_by_index.get(index_b) if index_b is not None else None
            _require(index_a is None or left_row is not None, f"row {physical_row}: A range")
            _require(index_b is None or right_row is not None, f"row {physical_row}: B range")
            for source in (left_row, right_row):
                if source is not None:
                    product_key = f"{source.county} {source.pm_full}".strip()
                    _require(
                        (source.route, product_key) == (route, visible_key),
                        f"row {physical_row}: displayed key/source drift",
                    )
            expected_status = (
                "Both" if left_row is not None and right_row is not None
                else f"{side_a} only" if left_row is not None else f"{side_b} only"
            )
            _require(values[5] == expected_status, f"row {physical_row}: status drift")
            state = values[13]
            _require(isinstance(state, str) and len(state) == 6, f"row {physical_row}: state")
            if left_row is not None and right_row is not None:
                _require(left_row.key == right_row.key, f"row {physical_row}: cross-key pair")
                expected_fields = _diff_fields(left_row, right_row, product=True)
                expected_state = "".join(
                    "N" if field in PRODUCT_CONTEXT_FIELDS
                    else "D" if field in expected_fields
                    else "E"
                    for field in PRODUCT_STATE_FIELDS
                )
                _require(state == expected_state, f"row {physical_row}: persisted state drift")
                _require(values[6] == len(expected_fields), f"row {physical_row}: diff count")
                pair = (index_a, index_b)
                _require(pair not in pairs, f"row {physical_row}: duplicate pair")
                pairs.add(pair)
                pair_diffs[pair] = expected_fields
            else:
                _require(state == "U" * 6 and values[6] is None, f"row {physical_row}: one-side state")
                if left_row is not None:
                    left_only.add(index_a)
                else:
                    assert index_b is not None
                    right_only.add(index_b)
            if index_a is not None:
                used_left[index_a] += 1
            if index_b is not None:
                used_right[index_b] += 1
            row_ids.append([route, visible_key, occurrence, index_a, index_b, state])
    finally:
        workbook.close()
        stream.close()

    _require(
        used_left == Counter(left_by_index.keys()),
        f"{path.name}: side-A coverage drift",
    )
    _require(
        used_right == Counter(right_by_index.keys()),
        f"{path.name}: side-B coverage drift",
    )
    calculated = _counts_for_pairs(pairs, left_by_index, right_by_index, product=True)
    workbook_counts = {
        "known": True,
        **calculated,
        "side_a_only_rows": len(left_only),
        "side_b_only_rows": len(right_only),
        "asserted_cells": len(pairs) * 3,
        "context_cells": len(pairs) * 3,
    }
    pair_records = [
        [a, b, list(pair_diffs[(a, b)])]
        for a, b in sorted(pairs, key=lambda pair: _pair_sort_key(pair, left_by_index))
    ]
    return {
        "captured_identity": captured_identity,
        "pairs": pairs,
        "left_only": left_only,
        "right_only": right_only,
        "pair_diffs": pair_diffs,
        "counts": workbook_counts,
        "ledger_identities": {
            "ordered_comparison_rows": _ledger(row_ids),
            "paired_asserted_states": _ledger(pair_records),
            "pair_map": _ledger([list(pair) for pair in sorted(pairs)]),
            "side_a_only": _ledger(sorted(left_only)),
            "side_b_only": _ledger(sorted(right_only)),
        },
    }


def _validate_product_result(
    label: str, spec: Mapping[str, object], scan: Mapping[str, object],
    parity: Mapping[str, object] | None,
) -> dict[str, object]:
    root = Path(spec["root"])
    result_path = root / "result.json"
    document, captured_result_identity = _captured_json(
        result_path, spec["result_binding"], f"{label} result",
    )
    _require(document.get("leg") == label, f"{label}: result leg drift")
    returned = document.get("result")
    _require(isinstance(returned, dict), f"{label}: result envelope missing")
    assert isinstance(returned, dict)
    _require(
        (returned.get("status"), returned.get("completion"), returned.get("pairing_quality"))
        == ("ok", "complete", "exact"),
        f"{label}: nonterminal result",
    )
    returned_counts = _normalize_returned_counts(returned.get("counts"), label)
    _require(returned_counts == scan["counts"], f"{label}: returned/workbook count drift")
    outputs = document.get("outputs")
    _require(isinstance(outputs, dict), f"{label}: output identities absent")
    assert isinstance(outputs, dict)
    authenticated_outputs = {}
    for flavor, filename in (("formulas", "comparison.xlsx"), ("values", "comparison (values).xlsx")):
        path = root / filename
        observed = (
            scan["captured_identity"] if flavor == "values" else _identity(path)
        )
        _require(outputs.get(flavor) == observed, f"{label}: {flavor} output/result drift")
        _require(_identity(path) == observed, f"{label}: {flavor} changed after capture")
        authenticated_outputs[flavor] = observed

    if parity is not None:
        parity_leg = parity.get("legs", {}).get(label)
        _require(isinstance(parity_leg, dict), f"{label}: absent from normalized parity")
        assert isinstance(parity_leg, dict)
        _require(
            parity_leg.get("publication", {}).get("persisted_counts") == scan["counts"],
            f"{label}: parity/workbook count drift",
        )
        _require(
            parity_leg.get("outputs") == authenticated_outputs,
            f"{label}: parity/output identity drift",
        )
        witness_result = parity.get("witness", {}).get("leg_results", {}).get(label, {}).get("result")
        _require(
            witness_result == captured_result_identity,
            f"{label}: parity/result identity drift",
        )

    return {
        "result": captured_result_identity,
        "outputs": authenticated_outputs,
        "returned_counts_rebuilt_exactly": True,
        "terminal_state": {
            "status": returned.get("status"),
            "completion": returned.get("completion"),
            "pairing_quality": returned.get("pairing_quality"),
        },
    }


def _validate_draft_leg(
    label: str, draft_leg: Mapping[str, object], pairing: Pairing,
    left: Sequence[Row], right: Sequence[Row],
) -> dict[str, object]:
    left_by_index = _row_by_index(left)
    right_by_index = _row_by_index(right)
    ordered_pairs = sorted(
        pairing.pairs, key=lambda pair: _pair_sort_key(pair, left_by_index)
    )
    all_counts: Counter[str] = Counter()
    asserted_counts: Counter[str] = Counter()
    all_rows = asserted_rows = 0
    expected_differing = draft_leg.get("differing_pairs")
    _require(isinstance(expected_differing, list), f"{label}: draft differing ledger absent")
    assert isinstance(expected_differing, list)
    differing_ordinal = 0
    for left_index, right_index in ordered_pairs:
        left_row = left_by_index[left_index]
        right_row = right_by_index[right_index]
        differing = list(_all_diff_fields(left_row, right_row))
        asserted = [field for field in differing if field in ASSERTED_FIELDS]
        if differing:
            all_rows += 1
            all_counts.update(differing)
            record = {
                "identity": list(left_row.key),
                "left_ref": left_row.source_ref,
                "right_ref": right_row.source_ref,
                "left_pm_full": left_row.pm_full,
                "right_pm_full": right_row.pm_full,
                "left_values": list(left_row.truth_values),
                "right_values": list(right_row.truth_values),
                "differing_fields": differing,
                "asserted_differing_fields": asserted,
                "left_kind": left_row.kind,
                "right_kind": right_row.kind,
            }
            _require(
                differing_ordinal < len(expected_differing)
                and expected_differing[differing_ordinal] == record,
                f"{label}: corrected draft differing pair {differing_ordinal} drift",
            )
            differing_ordinal += 1
        if asserted:
            asserted_rows += 1
            asserted_counts.update(asserted)
    _require(differing_ordinal == len(expected_differing), f"{label}: draft diff census drift")

    def one_sided(rows: Sequence[Row], indices: set[int]) -> list[dict[str, object]]:
        selected = sorted(
            (row for row in rows if row.product_index in indices),
            key=lambda row: (row.route, row.product_index),
        )
        return [{
            "ref": row.source_ref, "route": row.route, "county": row.county,
            "pm": row.pm_full, "values": list(row.truth_values), "kind": row.kind,
        } for row in selected]

    expected = {
        "left_rows": len(left),
        "right_rows": len(right),
        "paired_rows": len(pairing.pairs),
        "left_only_rows": len(pairing.left_only),
        "right_only_rows": len(pairing.right_only),
        "all_field_differing_rows": all_rows,
        "all_field_difference_cells": sum(all_counts.values()),
        "all_field_difference_counts": dict(sorted(all_counts.items())),
        "asserted_differing_rows": asserted_rows,
        "asserted_difference_cells": sum(asserted_counts.values()),
        "asserted_field_difference_counts": dict(sorted(asserted_counts.items())),
    }
    for key, value in expected.items():
        _require(draft_leg.get(key) == value, f"{label}: corrected draft {key} drift")
    _require(
        draft_leg.get("left_only") == one_sided(left, pairing.left_only),
        f"{label}: corrected draft left-only ledger drift",
    )
    _require(
        draft_leg.get("right_only") == one_sided(right, pairing.right_only),
        f"{label}: corrected draft right-only ledger drift",
    )
    draft_pairing = draft_leg.get("pairing")
    _require(isinstance(draft_pairing, dict), f"{label}: corrected pairing absent")
    assert isinstance(draft_pairing, dict)
    _require(
        draft_pairing.get("duplicate_traces") == pairing.duplicate_traces,
        f"{label}: corrected duplicate trace drift",
    )
    _require(
        draft_pairing.get("assignment_cost") == list(pairing.aggregate_cost),
        f"{label}: corrected aggregate assignment cost drift",
    )
    return {
        "counts": expected,
        "pair_map": _ledger([list(pair) for pair in sorted(pairing.pairs)]),
        "side_a_only": _ledger(sorted(pairing.left_only)),
        "side_b_only": _ledger(sorted(pairing.right_only)),
        "duplicate_traces": _ledger(pairing.duplicate_traces),
        "differing_pairs": _ledger(expected_differing),
        "corrected_draft_exact": True,
    }


def _projection_causes(row: Row) -> list[str]:
    if row.truth_values[-1] == row.product_values[-1]:
        return []
    causes: list[str] = []
    raw_without_xlsx_decode = _space(row.raw_description, xlsx=False)
    raw_with_xlsx_decode = _space(row.raw_description, xlsx=row.is_excel)
    prefix = ROUTE_PREFIX_RE.match(raw_with_xlsx_decode)
    if not row.is_tsmis and prefix is not None:
        causes.append("CMP-AUD-204_NUMERIC_TSN_PREFIX_DELETION")
    if row.is_tsmis and prefix is not None and _route(prefix.group(1)) != row.route:
        causes.append("CMP-AUD-204_SYMMETRIC_RULE_DELETES_TSMIS_CROSS_ROUTE_PREFIX")
    if (
        row.is_excel
        and raw_without_xlsx_decode != raw_with_xlsx_decode
        and OOXML_ESCAPE_RE.search(raw_without_xlsx_decode) is not None
    ):
        causes.append("CMP-AUD-197_UNDECODED_OOXML_CONTROL_ESCAPE")
    return causes


def _projection_population(rows: Sequence[Row]) -> dict[str, object]:
    records = []
    cause_counts: Counter[str] = Counter()
    unexplained = []
    for row in rows:
        truth = row.truth_values[-1]
        product = row.product_values[-1]
        if truth == product:
            continue
        causes = _projection_causes(row)
        record = {
            "product_index": row.product_index,
            "source_ref": row.source_ref,
            "identity": row.identity,
            "truth_description": truth,
            "product_description": product,
            "causes": causes,
        }
        records.append(record)
        cause_counts.update(causes)
        if not causes:
            unexplained.append(record)
    return {
        "rows": len(records),
        "cause_counts": dict(sorted(cause_counts.items())),
        "ledger": {**_ledger(records), "records": records},
        "unexplained": unexplained,
    }


def _pair_record(
    pair: tuple[int, int], left_by_index: Mapping[int, Row],
    right_by_index: Mapping[int, Row], *, product: bool,
) -> dict[str, object]:
    left = left_by_index[pair[0]]
    right = right_by_index[pair[1]]
    return {
        "side_a_index": pair[0],
        "side_b_index": pair[1],
        "identity": left.identity,
        "side_a_ref": left.source_ref,
        "side_b_ref": right.source_ref,
        "differing_fields": list(_diff_fields(left, right, product=product)),
    }


def _classify_fixed_pair_projection(
    label: str, pairing: Pairing, left: Sequence[Row], right: Sequence[Row],
    description_probe: Mapping[str, object],
) -> dict[str, object]:
    left_by_index = _row_by_index(left)
    right_by_index = _row_by_index(right)
    false_clean = []
    false_positive = []
    difference_preserving_mutations = []
    unexplained = []
    ordered_pairs = sorted(
        pairing.pairs, key=lambda pair: _pair_sort_key(pair, left_by_index)
    )
    for pair in ordered_pairs:
        left_row = left_by_index[pair[0]]
        right_row = right_by_index[pair[1]]
        truth_fields = set(_diff_fields(left_row, right_row, product=False))
        product_fields = set(_diff_fields(left_row, right_row, product=True))
        causes = sorted(set(
            _projection_causes(left_row) + _projection_causes(right_row)
        ))
        base = {
            "side_a_index": pair[0], "side_b_index": pair[1],
            "identity": left_row.identity,
            "side_a_ref": left_row.source_ref,
            "side_b_ref": right_row.source_ref,
            "truth_values": {
                "side_a": list(left_row.truth_values),
                "side_b": list(right_row.truth_values),
            },
            "product_values": {
                "side_a": list(left_row.product_values),
                "side_b": list(right_row.product_values),
            },
            "causes": causes,
        }
        for field in ASSERTED_FIELDS:
            if field in truth_fields and field not in product_fields:
                record = {**base, "field": field, "effect": "false_clean"}
                false_clean.append(record)
                if not causes:
                    unexplained.append(record)
            elif field not in truth_fields and field in product_fields:
                record = {**base, "field": field, "effect": "false_positive"}
                false_positive.append(record)
                if not causes:
                    unexplained.append(record)
        if truth_fields == product_fields and causes:
            difference_preserving_mutations.append({
                **base,
                "differing_fields_preserved": sorted(truth_fields),
                "effect": "source_claim_changed_without_aggregate_state_change",
            })

    false_clean_counts = Counter(
        cause for record in false_clean for cause in record["causes"]
    )
    false_positive_counts = Counter(
        cause for record in false_positive for cause in record["causes"]
    )
    cmp204_records = [{
        "identity": record["identity"],
        "tsmis": record["truth_values"]["side_a"][-1],
        "tsn": record["truth_values"]["side_b"][-1],
    } for record in false_clean if (
        record["field"] == "Description"
        and "CMP-AUD-204_NUMERIC_TSN_PREFIX_DELETION" in record["causes"]
    )]
    cmp204_content_sorted = sorted(
        cmp204_records,
        key=lambda item: (tuple(item["identity"]), item["tsmis"], item["tsn"]),
    )
    expected_leg = description_probe.get("comparison_legs", {}).get(label)
    _require(isinstance(expected_leg, dict), f"{label}: absent from Description probe")
    assert isinstance(expected_leg, dict)
    _require(
        len(cmp204_records) == expected_leg.get("product_false_clean_rows") == 81,
        f"{label}: CMP-AUD-204 false-clean census drift",
    )
    _require(
        _sha_bytes(_canonical(cmp204_content_sorted))
        == expected_leg.get("product_false_clean_content_sorted_sha256"),
        f"{label}: CMP-AUD-204 content ledger drift",
    )
    return {
        "oracle_pair_count": len(pairing.pairs),
        "oracle_truth_counts": _counts_for_pairs(
            pairing.pairs, left_by_index, right_by_index, product=False,
        ),
        "counterfactual_product_counts_on_oracle_pairs": _counts_for_pairs(
            pairing.pairs, left_by_index, right_by_index, product=True,
        ),
        "false_clean": {
            "rows": len(false_clean),
            "cause_counts": dict(sorted(false_clean_counts.items())),
            "ledger": {**_ledger(false_clean), "records": false_clean},
        },
        "false_positive": {
            "rows": len(false_positive),
            "cause_counts": dict(sorted(false_positive_counts.items())),
            "ledger": {**_ledger(false_positive), "records": false_positive},
        },
        "difference_preserving_source_claim_mutations": {
            **_ledger(difference_preserving_mutations),
            "records": difference_preserving_mutations,
        },
        "cmp_aud_204_content_sorted_sha256": _sha_bytes(
            _canonical(cmp204_content_sorted)
        ),
        "unexplained": unexplained,
    }


def _group_membership(
    pairing: Pairing, left_by_index: Mapping[int, Row],
    right_by_index: Mapping[int, Row],
) -> dict[tuple[str, str, str], dict[str, set[object]]]:
    result: dict[tuple[str, str, str], dict[str, set[object]]] = defaultdict(
        lambda: {"pairs": set(), "left_only": set(), "right_only": set()}
    )
    for pair in pairing.pairs:
        key = left_by_index[pair[0]].key
        _require(key == right_by_index[pair[1]].key, "pair crosses physical key")
        result[key]["pairs"].add(pair)
    for index in pairing.left_only:
        result[left_by_index[index].key]["left_only"].add(index)
    for index in pairing.right_only:
        result[right_by_index[index].key]["right_only"].add(index)
    return result


def _assignment_vector_for_pairs(
    pairs: set[tuple[int, int]], left: Sequence[Row], right: Sequence[Row],
) -> tuple[int, ...] | None:
    """Return the smaller-side vector or None when a purported map is malformed."""
    left_position = {row.product_index: index for index, row in enumerate(left)}
    right_position = {row.product_index: index for index, row in enumerate(right)}
    if len(pairs) != min(len(left), len(right)):
        return None
    seen_left: set[int] = set()
    seen_right: set[int] = set()
    for left_index, right_index in pairs:
        if (
            left_index not in left_position
            or right_index not in right_position
            or left_index in seen_left
            or right_index in seen_right
        ):
            return None
        seen_left.add(left_index)
        seen_right.add(right_index)
    if len(left) <= len(right):
        if seen_left != set(left_position):
            return None
        by_left = {left_index: right_position[right_index]
                   for left_index, right_index in pairs}
        return tuple(by_left[row.product_index] for row in left)
    if seen_right != set(right_position):
        return None
    by_right = {right_index: left_position[left_index]
                for left_index, right_index in pairs}
    return tuple(by_right[row.product_index] for row in right)


def _objective_total(
    pairs: set[tuple[int, int]], left_by_index: Mapping[int, Row],
    right_by_index: Mapping[int, Row], left_position: Mapping[int, int],
    right_position: Mapping[int, int],
    cost: Callable[[Row, Row, int, int], object], zero: object,
) -> object:
    total = zero
    for left_index, right_index in pairs:
        total = _add_cost(total, cost(
            left_by_index[left_index], right_by_index[right_index],
            left_position[left_index], right_position[right_index],
        ))
    return total


def _recomputed_group_membership(
    left: Sequence[Row], right: Sequence[Row], *, product: bool,
) -> tuple[dict[str, set[object]], object]:
    cost = _product_cost if product else _independent_cost
    zero: object = 0 if product else (0, 0, 0)
    pairs, left_only, right_only, total = _assignment(left, right, cost, zero)
    return ({
        "pairs": {(a.product_index, b.product_index) for a, b in pairs},
        "left_only": {row.product_index for row in left_only},
        "right_only": {row.product_index for row in right_only},
    }, total)


def _json_cost(value: object) -> object:
    return list(value) if isinstance(value, tuple) else value


def _assignment_residuals(
    independent: Pairing, product: Pairing,
    left: Sequence[Row], right: Sequence[Row],
    *, cmp159_normalized_index: int | None,
) -> dict[str, object]:
    left_by_index = _row_by_index(left)
    right_by_index = _row_by_index(right)
    independent_groups = _group_membership(independent, left_by_index, right_by_index)
    product_groups = _group_membership(product, left_by_index, right_by_index)
    records = []
    unexplained = []
    aggregate_field_delta: Counter[str] = Counter()
    aggregate_row_delta = 0
    ft_records = []
    left_groups: dict[tuple[str, str, str], list[Row]] = defaultdict(list)
    right_groups: dict[tuple[str, str, str], list[Row]] = defaultdict(list)
    for row in left:
        left_groups[row.key].append(row)
    for row in right:
        right_groups[row.key].append(row)
    for identity in sorted(set(independent_groups) | set(product_groups)):
        independent_group = independent_groups.get(
            identity, {"pairs": set(), "left_only": set(), "right_only": set()}
        )
        product_group = product_groups.get(
            identity, {"pairs": set(), "left_only": set(), "right_only": set()}
        )
        if independent_group == product_group:
            continue
        oracle_pairs = independent_group["pairs"]
        persisted_pairs = product_group["pairs"]
        oracle_counts = _counts_for_pairs(
            oracle_pairs, left_by_index, right_by_index, product=True,
        )
        persisted_counts = _counts_for_pairs(
            persisted_pairs, left_by_index, right_by_index, product=True,
        )
        field_delta = {
            field: (
                persisted_counts["per_field_counts"][field]
                - oracle_counts["per_field_counts"][field]
            ) for field in ASSERTED_FIELDS
        }
        row_delta = persisted_counts["differing_rows"] - oracle_counts["differing_rows"]
        aggregate_field_delta.update(field_delta)
        aggregate_row_delta += row_delta
        group_left = left_groups.get(identity, [])
        group_right = right_groups.get(identity, [])
        group_rows = [*group_left, *group_right]
        expected_source, expected_source_total = _recomputed_group_membership(
            group_left, group_right, product=False,
        )
        expected_product, expected_product_total = _recomputed_group_membership(
            group_left, group_right, product=True,
        )
        source_vector = _assignment_vector_for_pairs(
            independent_group["pairs"], group_left, group_right,
        )
        product_vector = _assignment_vector_for_pairs(
            product_group["pairs"], group_left, group_right,
        )
        source_exact = independent_group == expected_source
        product_exact = product_group == expected_product
        source_preference = product_preference = False
        source_selected_total = source_alternative_total = None
        product_selected_total = product_alternative_total = None
        causes: list[str] = []
        reasons: list[str] = []
        if not source_exact:
            reasons.append("SOURCE_PAIRING_IS_NOT_RECOMPUTED_SOURCE_OPTIMUM")
        if not product_exact:
            reasons.append("PRODUCT_PAIRING_IS_NOT_RECOMPUTED_PRODUCT_OPTIMUM")
        if source_vector is None:
            reasons.append("SOURCE_PAIR_MAP_IS_NOT_A_COMPLETE_RECTANGULAR_ASSIGNMENT")
        if product_vector is None:
            reasons.append("PRODUCT_PAIR_MAP_IS_NOT_A_COMPLETE_RECTANGULAR_ASSIGNMENT")
        if source_vector is not None and product_vector is not None:
            left_position = {
                row.product_index: index for index, row in enumerate(group_left)
            }
            right_position = {
                row.product_index: index for index, row in enumerate(group_right)
            }
            source_selected_total = _objective_total(
                independent_group["pairs"], left_by_index, right_by_index,
                left_position, right_position, _independent_cost, (0, 0, 0),
            )
            source_alternative_total = _objective_total(
                product_group["pairs"], left_by_index, right_by_index,
                left_position, right_position, _independent_cost, (0, 0, 0),
            )
            product_selected_total = _objective_total(
                product_group["pairs"], left_by_index, right_by_index,
                left_position, right_position, _product_cost, 0,
            )
            product_alternative_total = _objective_total(
                independent_group["pairs"], left_by_index, right_by_index,
                left_position, right_position, _product_cost, 0,
            )
            if source_selected_total < source_alternative_total:
                source_preference = True
                causes.append("SOURCE_ALL_FIELD_OBJECTIVE_STRICTLY_PREFERS_ORACLE")
            elif (
                source_selected_total == source_alternative_total
                and source_vector < product_vector
            ):
                source_preference = True
                causes.append("SOURCE_LEXICOGRAPHIC_TIEBREAK_PREFERS_ORACLE")
            else:
                reasons.append("SOURCE_SELECTION_PREFERENCE_NOT_PROVEN")
            if product_selected_total < product_alternative_total:
                product_preference = True
                causes.append("PRODUCT_ASSERTED_OBJECTIVE_STRICTLY_PREFERS_PERSISTED")
            elif (
                product_selected_total == product_alternative_total
                and product_vector < source_vector
            ):
                product_preference = True
                causes.append("PRODUCT_LEXICOGRAPHIC_TIEBREAK_PREFERS_PERSISTED")
            else:
                reasons.append("PRODUCT_SELECTION_PREFERENCE_NOT_PROVEN")
        policy_proven = (
            source_exact and product_exact
            and source_preference and product_preference
            and not reasons
        )
        if policy_proven:
            causes.append(
                "PRODUCT_ASSERTED_ONLY_VS_SOURCE_ALL_FIELDS_ASSIGNMENT_POLICY"
            )
        projection_causes = sorted(set(
            cause for row in group_rows for cause in _projection_causes(row)
        ))
        causes.extend(projection_causes)
        if cmp159_normalized_index is not None and any(
            not row.is_tsmis and row.product_index == cmp159_normalized_index
            for row in group_rows
        ):
            causes.append("CMP-AUD-159_NORMALIZED_PUNCTUATION_INFLUENCES_ASSIGNMENT")
        selection = {
            "source_pairing_recomputed_exact": source_exact,
            "product_pairing_recomputed_exact": product_exact,
            "source_assignment_vector": (
                list(source_vector) if source_vector is not None else None
            ),
            "product_assignment_vector": (
                list(product_vector) if product_vector is not None else None
            ),
            "source_selected_source_objective": _json_cost(source_selected_total),
            "product_alternative_source_objective": _json_cost(source_alternative_total),
            "product_selected_product_objective": product_selected_total,
            "source_alternative_product_objective": product_alternative_total,
            "recomputed_source_optimum": _json_cost(expected_source_total),
            "recomputed_product_optimum": expected_product_total,
            "source_selection_preference_proven": source_preference,
            "product_selection_preference_proven": product_preference,
            "policy_difference_executably_proven": policy_proven,
        }
        record = {
            "identity": list(identity),
            "causes": causes,
            "attribution": selection,
            "oracle": {
                "pairs": [
                    _pair_record(pair, left_by_index, right_by_index, product=True)
                    for pair in sorted(oracle_pairs)
                ],
                "side_a_only": sorted(independent_group["left_only"]),
                "side_b_only": sorted(independent_group["right_only"]),
                "counterfactual_product_counts": oracle_counts,
            },
            "persisted_product": {
                "pairs": [
                    _pair_record(pair, left_by_index, right_by_index, product=True)
                    for pair in sorted(persisted_pairs)
                ],
                "side_a_only": sorted(product_group["left_only"]),
                "side_b_only": sorted(product_group["right_only"]),
                "counts": persisted_counts,
            },
            "aggregate_contribution": {
                "per_field_counts": field_delta,
                "differing_rows": row_delta,
                "differing_cells": sum(field_delta.values()),
            },
        }
        records.append(record)
        if reasons:
            unexplained.append({
                "identity": list(identity),
                "reasons": reasons,
                "attribution": selection,
                "oracle_pairs": [list(pair) for pair in sorted(oracle_pairs)],
                "persisted_pairs": [list(pair) for pair in sorted(persisted_pairs)],
                "expected_product_pairs": [
                    list(pair) for pair in sorted(expected_product["pairs"])
                ],
            })
        if field_delta["FT"]:
            ft_records.append(record)
    all_attributed = (
        not unexplained
        and all(
            record["attribution"]["policy_difference_executably_proven"]
            for record in records
        )
    )
    return {
        "changed_groups": len(records),
        "pair_additions": len(product.pairs - independent.pairs),
        "pair_removals": len(independent.pairs - product.pairs),
        "side_a_membership_additions": len(product.left_only - independent.left_only),
        "side_a_membership_removals": len(independent.left_only - product.left_only),
        "side_b_membership_additions": len(product.right_only - independent.right_only),
        "side_b_membership_removals": len(independent.right_only - product.right_only),
        "aggregate_contribution": {
            "per_field_counts": {
                field: aggregate_field_delta[field] for field in ASSERTED_FIELDS
            },
            "differing_rows": aggregate_row_delta,
            "differing_cells": sum(aggregate_field_delta.values()),
        },
        "ledger": {**_ledger(records), "records": records},
        "ft_change_ledger": {**_ledger(ft_records), "records": ft_records},
        "unexplained": {**_ledger(unexplained), "records": unexplained},
        "all_pair_map_differences_have_recomputed_policy_cause": all_attributed,
    }


def _assignment_attribution_mutations() -> dict[str, object]:
    """Prove authentic policy divergence passes and an arbitrary swap fails."""
    def synthetic(dataset: str, index: int, city: str, is_tsmis: bool) -> Row:
        return Row(
            dataset=dataset,
            product_index=index,
            source_ref=f"{dataset}:{index}",
            route="001",
            county="AAA",
            pm_base="001.000",
            pm_suffix="",
            values=(city, "", "", "", "SAME"),
            raw_description="SAME",
            kind="data",
            is_tsmis=is_tsmis,
            is_excel=False,
        )

    left = [
        synthetic("synthetic_tsmis", index, city, True)
        for index, city in enumerate(("A", "B", "C"))
    ]
    right = [
        synthetic("synthetic_tsn", index, city, False)
        for index, city in enumerate(("B", "C", "A"))
    ]
    source_pairing = _pair_rows(left, right, product=False)
    authentic_product = _pair_rows(left, right, product=True)
    _require(
        source_pairing.pairs == {(0, 2), (1, 0), (2, 1)}
        and authentic_product.pairs == {(0, 0), (1, 1), (2, 2)},
        "assignment attribution fixture no longer exercises divergent policies",
    )
    authentic = _assignment_residuals(
        source_pairing, authentic_product, left, right,
        cmp159_normalized_index=None,
    )
    _require(
        authentic["changed_groups"] == 1
        and authentic["unexplained"]["rows"] == 0
        and authentic["all_pair_map_differences_have_recomputed_policy_cause"] is True,
        "authentic assignment policy divergence was not attributed",
    )

    arbitrary_pairs = {(0, 1), (1, 0), (2, 2)}
    arbitrary = Pairing(arbitrary_pairs, set(), set(), [], 0)
    rejected = _assignment_residuals(
        source_pairing, arbitrary, left, right,
        cmp159_normalized_index=None,
    )
    _require(
        rejected["changed_groups"] == 1
        and rejected["unexplained"]["rows"] == 1
        and rejected["all_pair_map_differences_have_recomputed_policy_cause"] is False
        and "PRODUCT_PAIRING_IS_NOT_RECOMPUTED_PRODUCT_OPTIMUM"
        in rejected["unexplained"]["records"][0]["reasons"],
        "arbitrary assignment swap did not fail attribution",
    )
    return {
        "authentic_policy_divergence": {
            "status": "passed",
            "source_pair_map": _ledger(
                [list(pair) for pair in sorted(source_pairing.pairs)]
            ),
            "product_pair_map": _ledger(
                [list(pair) for pair in sorted(authentic_product.pairs)]
            ),
            "attribution": _ledger(authentic["ledger"]["records"]),
        },
        "arbitrary_swap": {
            "status": "rejected",
            "pair_map": _ledger([list(pair) for pair in sorted(arbitrary_pairs)]),
            "unexplained": _ledger(rejected["unexplained"]["records"]),
            "reasons": rejected["unexplained"]["records"][0]["reasons"],
        },
    }


def _reconciliation(
    oracle: Mapping[str, object], fixed_product: Mapping[str, object],
    persisted: Mapping[str, object], assignment: Mapping[str, object],
) -> dict[str, object]:
    fields = {}
    all_exact = True
    assignment_fields = assignment["aggregate_contribution"]["per_field_counts"]
    for field in ASSERTED_FIELDS:
        truth = oracle["per_field_counts"][field]
        fixed = fixed_product["per_field_counts"][field]
        actual = persisted["per_field_counts"].get(field, 0)
        projection_delta = fixed - truth
        assignment_delta = assignment_fields[field]
        exact = truth + projection_delta + assignment_delta == actual
        all_exact &= exact
        fields[field] = {
            "oracle_truth": truth,
            "projection_delta": projection_delta,
            "counterfactual_product_on_oracle_pairs": fixed,
            "assignment_delta": assignment_delta,
            "persisted_product": actual,
            "exact": exact,
        }
    truth_rows = oracle["differing_rows"]
    fixed_rows = fixed_product["differing_rows"]
    actual_rows = persisted["differing_rows"]
    assignment_rows = assignment["aggregate_contribution"]["differing_rows"]
    row_exact = truth_rows + (fixed_rows - truth_rows) + assignment_rows == actual_rows
    all_exact &= row_exact
    side_a_exact = persisted["side_a_only_rows"] >= 0
    side_b_exact = persisted["side_b_only_rows"] >= 0
    return {
        "fields": fields,
        "differing_rows": {
            "oracle_truth": truth_rows,
            "projection_delta": fixed_rows - truth_rows,
            "counterfactual_product_on_oracle_pairs": fixed_rows,
            "assignment_delta": assignment_rows,
            "persisted_product": actual_rows,
            "exact": row_exact,
        },
        "differing_cells": {
            "oracle_truth": oracle["differing_cells"],
            "projection_delta": fixed_product["differing_cells"] - oracle["differing_cells"],
            "assignment_delta": assignment["aggregate_contribution"]["differing_cells"],
            "persisted_product": persisted["differing_cells"],
            "exact": sum(item["persisted_product"] for item in fields.values())
            == persisted["differing_cells"],
        },
        "side_counts_present": side_a_exact and side_b_exact,
        "all_aggregate_deltas_reconciled": all_exact,
    }


def _raw_normalized_ledger(
    raw_known: Sequence[Row], normalized: Sequence[Row],
) -> dict[str, object]:
    _require(len(raw_known) == len(normalized), "raw/normalized known-row census drift")
    fields = ("Route", "County", "PM", *VALUE_FIELDS)
    records = []
    counts: Counter[str] = Counter()
    for ordinal, (raw, norm) in enumerate(zip(raw_known, normalized, strict=True), 1):
        raw_values = (raw.route, raw.county, raw.pm_full, *raw.truth_values)
        normalized_values = (norm.route, norm.county, norm.pm_full, *norm.truth_values)
        differing = [
            field for field, a, b
            in zip(fields, raw_values, normalized_values, strict=True)
            if a != b
        ]
        _require(raw.key == norm.key, f"raw/normalized identity drift at {ordinal}")
        if differing:
            counts.update(differing)
            records.append({
                "ordinal": ordinal,
                "raw_ref": raw.source_ref,
                "normalized_ref": norm.source_ref,
                "raw_values": list(raw_values),
                "normalized_values": list(normalized_values),
                "differing_fields": differing,
            })
    return {
        "rows": len(raw_known),
        "differing_rows": len(records),
        "difference_cells": sum(counts.values()),
        "field_difference_counts": dict(sorted(counts.items())),
        "differences": records,
    }


def _counts_delta(
    later: Mapping[str, object], earlier: Mapping[str, object],
) -> dict[str, object]:
    return {
        "differing_rows": later["differing_rows"] - earlier["differing_rows"],
        "differing_cells": later["differing_cells"] - earlier["differing_cells"],
        "per_field_counts": {
            field: (
                later["per_field_counts"].get(field, 0)
                - earlier["per_field_counts"].get(field, 0)
            ) for field in ASSERTED_FIELDS
        },
    }


def _cross_form_overlay(
    runtime_legs: Mapping[str, Mapping[str, object]],
    raw_known: Sequence[Row], normalized: Sequence[Row],
    raw_normalized: Mapping[str, object],
) -> dict[str, object]:
    description_records = [
        item for item in raw_normalized["differences"]
        if "Description" in item["differing_fields"]
    ]
    _require(len(description_records) == 1, "CMP-AUD-159 punctuation census drift")
    punctuation = description_records[0]
    normalized_index = punctuation["ordinal"] - 1
    raw_index = raw_known[normalized_index].product_index
    _require(
        punctuation["raw_values"][-1]
        == "KEMWATER CHEMICAL PLANT - RT/FRONTAGE ROAD - LT."
        and punctuation["normalized_values"][-1]
        == "KEMWATER CHEMICAL PLANT - RT/FRONTAGE, ROAD - LT.",
        "CMP-AUD-159 exact punctuation record drift",
    )
    _require(
        raw_normalized["field_difference_counts"]
        == {"Description": 1, "Distance To Next Point": 565},
        "raw/normalized projection census drift",
    )

    by_tsmis = {}
    for prefix in ("excel", "pdf"):
        raw_leg = runtime_legs[f"{prefix}_vs_raw_tsn"]
        norm_leg = runtime_legs[f"{prefix}_vs_normalized_tsn"]
        raw_product_pairs = raw_leg["product_pairing"].pairs
        norm_product_pairs = norm_leg["product_pairing"].pairs
        raw_oracle_pairs = raw_leg["independent_pairing"].pairs
        norm_oracle_pairs = norm_leg["independent_pairing"].pairs
        mapped_norm_product = {
            (left_index, raw_known[right_index].product_index)
            for left_index, right_index in norm_product_pairs
        }
        mapped_norm_oracle = {
            (left_index, raw_known[right_index].product_index)
            for left_index, right_index in norm_oracle_pairs
        }
        raw_target_product = sorted(
            pair for pair in raw_product_pairs if pair[1] == raw_index
        )
        norm_target_product = sorted(
            pair for pair in norm_product_pairs if pair[1] == normalized_index
        )
        _require(
            len(raw_target_product) == len(norm_target_product) == 1,
            f"{prefix}: CMP-AUD-159 row is not paired exactly once",
        )
        raw_pair = raw_target_product[0]
        norm_pair = norm_target_product[0]
        raw_left = raw_leg["left_by_index"][raw_pair[0]]
        norm_left = norm_leg["left_by_index"][norm_pair[0]]
        raw_right = raw_leg["right_by_index"][raw_pair[1]]
        norm_right = norm_leg["right_by_index"][norm_pair[1]]
        raw_fields = _diff_fields(raw_left, raw_right, product=True)
        norm_fields = _diff_fields(norm_left, norm_right, product=True)
        _require(
            "Description" not in raw_fields and "Description" in norm_fields,
            f"{prefix}: CMP-AUD-159 did not create exactly the product false positive",
        )
        by_tsmis[prefix] = {
            "cmp_aud_159_product_false_positive": {
                "raw_pair": _pair_record(
                    raw_pair, raw_leg["left_by_index"], raw_leg["right_by_index"],
                    product=True,
                ),
                "normalized_pair": _pair_record(
                    norm_pair, norm_leg["left_by_index"], norm_leg["right_by_index"],
                    product=True,
                ),
                "raw_description": punctuation["raw_values"][-1],
                "normalized_description": punctuation["normalized_values"][-1],
                "effect": "normalized product adds one false-positive Description cell",
            },
            "product_pair_map_raw_vs_normalized": {
                "additions": len(mapped_norm_product - raw_product_pairs),
                "removals": len(raw_product_pairs - mapped_norm_product),
                "raw_pair_map": _ledger([list(pair) for pair in sorted(raw_product_pairs)]),
                "mapped_normalized_pair_map": _ledger(
                    [list(pair) for pair in sorted(mapped_norm_product)]
                ),
            },
            "oracle_pair_map_raw_vs_normalized": {
                "additions": len(mapped_norm_oracle - raw_oracle_pairs),
                "removals": len(raw_oracle_pairs - mapped_norm_oracle),
                "cause": (
                    "565 CMP-AUD-156 pointer projections plus the one CMP-AUD-159 "
                    "Description projection alter the all-field source assignment objective"
                ),
            },
            "product_count_delta_normalized_minus_raw": _counts_delta(
                norm_leg["scan"]["counts"], raw_leg["scan"]["counts"]
            ),
            "oracle_count_delta_normalized_minus_raw": _counts_delta(
                norm_leg["fixed_projection"]["oracle_truth_counts"],
                raw_leg["fixed_projection"]["oracle_truth_counts"],
            ),
        }
    return {
        "raw_to_normalized_projection": {
            "counts": {
                key: raw_normalized[key]
                for key in ("rows", "differing_rows", "difference_cells", "field_difference_counts")
            },
            "ledger": {**_ledger(raw_normalized["differences"]),
                       "records": raw_normalized["differences"]},
        },
        "cmp_aud_159": punctuation,
        "per_tsmis_form": by_tsmis,
    }


def _write_atomic(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent,
    )
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    except BaseException:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise


def run(output: Path) -> dict[str, object]:
    protected_paths = _bound_artifact_paths()
    output = _guard_output_path(output, protected_paths)
    path_and_alias_mutations = _path_and_alias_mutation_probes()
    assignment_attribution_mutations = _assignment_attribution_mutations()
    inputs_before = _bind_inputs()
    captured_json_identities: dict[str, dict[str, object]] = {}

    def captured(label: str) -> dict[str, object]:
        path, expected_bytes, expected_sha = INPUT_BINDINGS[label]
        document, identity = _captured_json(
            path, {"bytes": expected_bytes, "sha256": expected_sha}, label,
        )
        _require(identity == inputs_before[label], f"{label}: guard/capture identity drift")
        captured_json_identities[label] = identity
        return document

    source_document = captured("source_rows")
    tsn_document = captured("tsn_rows")
    source_oracle = captured("source_oracle")
    draft = captured("corrected_comparison_draft")
    description_probe = captured("description_normalization_probe")
    excel_escape_probe = captured("installed_excel_escape_probe")
    parity = captured("normalized_product_parity")

    _require(source_document.get("not_an_acceptance_artifact") is True,
             "source cache lost non-acceptance marker")
    _require(tsn_document.get("not_an_acceptance_artifact") is True,
             "TSN cache lost non-acceptance marker")
    _require(draft.get("not_an_acceptance_artifact") is True,
             "comparison draft lost non-acceptance marker")
    _require(source_oracle.get("audit_complete") is False,
             "source-oracle lifecycle unexpectedly changed")
    _require(source_oracle.get("status") == "source-parsed",
             "source-oracle status drift")
    _require(
        source_oracle.get("extraction_invariants", {}).get("current_excel_rows_60494")
        and source_oracle.get("extraction_invariants", {}).get("current_pdf_rows_60493"),
        "source-oracle current census invariants absent",
    )
    _require(
        description_probe.get("artifact_status") == "NON_ACCEPTANCE_DEVELOPMENT_PROBE"
        and description_probe.get("invariants", {}).get("all_checks_passed") is True,
        "Description probe lifecycle/invariants drift",
    )
    _require(
        excel_escape_probe.get("workbook_opened_read_only") is True
        and len(excel_escape_probe.get("cells", [])) == 4
        and all(excel_escape_probe.get("invariants", {}).values()),
        "installed-Excel OOXML proof drift",
    )
    _require(
        parity.get("status") == "pass_with_expected_product_defect"
        and parity.get("acceptance_artifact") is False,
        "normalized product parity lifecycle drift",
    )

    serialized_sources = source_document.get("rows")
    _require(isinstance(serialized_sources, dict), "source row cache envelope drift")
    assert isinstance(serialized_sources, dict)
    excel = _load_tsmis(serialized_sources["current_tsmis_excel"], "current_tsmis_excel")
    pdf = _load_tsmis(serialized_sources["current_tsmis_pdf"], "current_tsmis_pdf")
    raw_all, raw_known, raw_unknown = _load_raw_tsn(tsn_document["raw_records"])
    normalized = _load_normalized_tsn(tsn_document["normalized"]["rows"])
    _require(
        (len(excel), len(pdf), len(raw_all), len(raw_known), len(raw_unknown), len(normalized))
        == (60_494, 60_493, 69_804, 69_758, 46, 69_758),
        "source census drift",
    )
    datasets = {
        "current_tsmis_excel": excel,
        "current_tsmis_pdf": pdf,
        "raw_tsn": raw_all,
        "raw_tsn_known_county": raw_known,
        "normalized_tsn": normalized,
    }

    raw_normalized = _raw_normalized_ledger(raw_known, normalized)
    _require(
        raw_normalized == draft.get("raw_vs_normalized_tsn"),
        "raw/normalized independent ledger drift from corrected draft",
    )
    cmp159_records = [
        record for record in raw_normalized["differences"]
        if "Description" in record["differing_fields"]
    ]
    _require(len(cmp159_records) == 1, "CMP-AUD-159 exact record missing")
    cmp159_normalized_index = cmp159_records[0]["ordinal"] - 1

    projection_populations = {
        label: _projection_population(rows)
        for label, rows in (
            ("current_tsmis_excel", excel),
            ("current_tsmis_pdf", pdf),
            ("raw_tsn", raw_all),
            ("normalized_tsn", normalized),
        )
    }
    _require(
        projection_populations["raw_tsn"]["rows"]
        == projection_populations["normalized_tsn"]["rows"] == 154,
        "CMP-AUD-204 TSN projection census drift",
    )

    runtime_legs: dict[str, dict[str, object]] = {}
    output_legs: dict[str, dict[str, object]] = {}
    unexplained: list[dict[str, object]] = []
    unknown_indices = {row.product_index for row in raw_unknown}
    for label, spec in LEG_SPECS.items():
        left = datasets[str(spec["left_dataset"])]
        right_product = datasets[str(spec["right_dataset"])]
        right_independent = (
            datasets["raw_tsn_known_county"]
            if spec["right_dataset"] == "raw_tsn" else right_product
        )
        independent = _pair_rows(left, right_independent, product=False)
        draft_leg = draft.get("tsmis_vs_tsn", {}).get(spec["draft_leg"])
        _require(isinstance(draft_leg, dict), f"{label}: corrected draft leg absent")
        assert isinstance(draft_leg, dict)
        draft_validation = _validate_draft_leg(
            label, draft_leg, independent, left, right_independent,
        )

        product_pairing = _pair_rows(left, right_product, product=True)
        values_path = Path(spec["root"]) / "comparison (values).xlsx"
        scan = _scan_product_workbook(
            values_path, str(spec["side_a"]), str(spec["side_b"]),
            left, right_product, spec["values_binding"],
        )
        _require(scan["pairs"] == product_pairing.pairs,
                 f"{label}: persisted/recomputed product pair map drift")
        _require(scan["left_only"] == product_pairing.left_only,
                 f"{label}: persisted/recomputed product side-A inventory drift")
        _require(scan["right_only"] == product_pairing.right_only,
                 f"{label}: persisted/recomputed product side-B inventory drift")
        product_authentication = _validate_product_result(
            label, spec, scan, parity if "normalized" in label else None,
        )

        fixed_projection = _classify_fixed_pair_projection(
            label, independent, left, right_independent, description_probe,
        )
        comparable_product = product_pairing
        unknown_ledger = []
        if spec["right_dataset"] == "raw_tsn":
            _require(
                unknown_indices <= product_pairing.right_only,
                f"{label}: blank-County raw TSN rows were not all product one-sided",
            )
            comparable_product = Pairing(
                set(product_pairing.pairs), set(product_pairing.left_only),
                set(product_pairing.right_only) - unknown_indices,
                product_pairing.duplicate_traces, product_pairing.aggregate_cost,
            )
            unknown_ledger = [{
                "product_index": row.product_index,
                "source_ref": row.source_ref,
                "identity": row.identity,
                "description": row.truth_values[-1],
                "classification": "CMP-AUD-158_EXPLICIT_UNKNOWN_COUNTY_TSN_ONLY",
            } for row in raw_unknown]

        assignment = _assignment_residuals(
            independent, comparable_product, left, right_product,
            cmp159_normalized_index=(
                cmp159_normalized_index if spec["right_dataset"] == "normalized_tsn"
                else None
            ),
        )
        reconciliation = _reconciliation(
            fixed_projection["oracle_truth_counts"],
            fixed_projection["counterfactual_product_counts_on_oracle_pairs"],
            scan["counts"], assignment,
        )
        _require(
            reconciliation["all_aggregate_deltas_reconciled"],
            f"{label}: aggregate delta failed exact reconciliation",
        )

        side_a_expected = (
            len(independent.left_only)
            + assignment["side_a_membership_additions"]
            - assignment["side_a_membership_removals"]
        )
        side_b_expected_known = (
            len(independent.right_only)
            + assignment["side_b_membership_additions"]
            - assignment["side_b_membership_removals"]
        )
        _require(side_a_expected == scan["counts"]["side_a_only_rows"],
                 f"{label}: side-A count reconciliation drift")
        _require(
            side_b_expected_known + len(unknown_ledger)
            == scan["counts"]["side_b_only_rows"],
            f"{label}: side-B count reconciliation drift",
        )
        unexplained.extend({"leg": label, **record}
                            for record in fixed_projection["unexplained"])
        unexplained.extend({
            "leg": label,
            "classification": "unexplained_assignment_attribution",
            **record,
        } for record in assignment["unexplained"]["records"])

        runtime_legs[label] = {
            "left_by_index": _row_by_index(left),
            "right_by_index": _row_by_index(right_product),
            "independent_pairing": independent,
            "product_pairing": product_pairing,
            "scan": scan,
            "fixed_projection": fixed_projection,
        }
        output_legs[label] = {
            "product_authentication": product_authentication,
            "persisted_product": {
                "counts": scan["counts"],
                "ledgers": scan["ledger_identities"],
                "recomputed_product_assignment_exact": True,
            },
            "independent_oracle": draft_validation,
            "fixed_oracle_pair_projection_effects": fixed_projection,
            "assignment_policy_effects": assignment,
            "raw_unknown_county": {
                "rows": len(unknown_ledger),
                "ledger": {**_ledger(unknown_ledger), "records": unknown_ledger},
            },
            "reconciliation": reconciliation,
            "side_count_reconciliation": {
                "side_a_expected": side_a_expected,
                "side_a_persisted": scan["counts"]["side_a_only_rows"],
                "side_b_known_expected": side_b_expected_known,
                "side_b_unknown_county": len(unknown_ledger),
                "side_b_persisted": scan["counts"]["side_b_only_rows"],
                "exact": True,
            },
        }

    overlay = _cross_form_overlay(
        runtime_legs, raw_known, normalized, raw_normalized,
    )
    for dataset, population in projection_populations.items():
        unexplained.extend(
            {"dataset": dataset, "classification": "unexplained_source_projection", **record}
            for record in population["unexplained"]
        )

    inputs_after = _bind_inputs()
    _require(inputs_after == inputs_before, "identity-bound input changed during run")
    output = _guard_output_path(output, protected_paths)
    result: dict[str, object] = {
        "audit": "Stage 8 Highway Sequence four-leg product residual classifier",
        "artifact_status": "NON_ACCEPTANCE_DEVELOPMENT_CLASSIFIER",
        "acceptance_eligible": False,
        "not_an_acceptance_artifact": (
            "This classifier consumes frozen development row caches and already-published "
            "workbooks. Final acceptance must independently reparse immutable raw sources."
        ),
        "inputs_before": inputs_before,
        "inputs_after": inputs_after,
        "captured_json_identities": captured_json_identities,
        "guard_mutations": {
            "path_and_output_alias": path_and_alias_mutations,
            "assignment_attribution": assignment_attribution_mutations,
        },
        "source_census": {
            "current_tsmis_excel": len(excel),
            "current_tsmis_pdf": len(pdf),
            "raw_tsn_all": len(raw_all),
            "raw_tsn_known_county": len(raw_known),
            "raw_tsn_unknown_county": len(raw_unknown),
            "normalized_tsn": len(normalized),
        },
        "source_projection_populations": projection_populations,
        "legs": output_legs,
        "raw_normalized_overlay": overlay,
        "unexplained_residuals": unexplained,
        "methodology": {
            "product_aggregate_counts_trusted": False,
            "persisted_comparison_rows_read": True,
            "product_pairing_recomputed_without_product_imports": True,
            "independent_pairing_recomputed_from_bound_caches": True,
            "corrected_draft_differences_one_by_one_verified": True,
            "json_parsed_from_hashed_captured_bytes": True,
            "comparison_workbooks_parsed_from_hashed_captured_bytes": True,
            "filesystem_pre_and_post_guards_equal_captured_identities": True,
            "supplied_paths_checked_for_symlink_reparse_before_resolution": True,
            "output_checked_against_every_bound_artifact_before_and_after_run": True,
            "assignment_attribution_is_executable_and_mutation_tested": True,
            "classification_order": [
                "source-fixed pair normalization effects",
                "product-vs-source assignment policy effects",
                "raw unknown-County visibility",
                "raw-vs-normalized projection overlay",
            ],
        },
        "invariants": {
            "all_input_identities_stable": True,
            "all_four_product_pair_maps_recomputed_exactly": True,
            "all_four_corrected_source_pair_maps_recomputed_exactly": True,
            "all_four_aggregate_deltas_reconcile": True,
            "all_four_assignment_policy_attributions_executably_proven": all(
                leg["assignment_policy_effects"]
                ["all_pair_map_differences_have_recomputed_policy_cause"]
                for leg in output_legs.values()
            ),
            "authentic_and_arbitrary_swap_mutations_passed": True,
            "output_alias_mutations_passed_or_truthfully_skipped": True,
            "cmp_aud_204_false_clean_rows_each_leg": 81,
            "cmp_aud_159_normalized_false_positive_each_tsmis_form": 1,
            "raw_unknown_county_rows_explicit": 46,
            "unexplained_residuals": len(unexplained),
            "all_residuals_classified": not unexplained,
        },
    }
    payload = _canonical(result, newline=True)
    _write_atomic(output, payload)
    print(json.dumps({
        "status": "PASS" if not unexplained else "UNEXPLAINED_RESIDUALS",
        "output": str(output.resolve()),
        "bytes": len(payload),
        "sha256": _sha_bytes(payload),
        "unexplained_residuals": len(unexplained),
        "legs": {
            label: output_legs[label]["reconciliation"]["fields"]
            for label in output_legs
        },
    }, ensure_ascii=False, sort_keys=True))
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    arguments = parser.parse_args()
    run(arguments.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
