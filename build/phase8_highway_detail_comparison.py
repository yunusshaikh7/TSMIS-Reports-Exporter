#!/usr/bin/env python3
"""Independent Stage-8 Highway Detail source/comparison oracle.

The source layer imports no application parser, consolidator, comparator,
schema, evidence adapter, or workbook result.  Current TSMIS PDF rows come from
``phase8_highway_detail_source_oracle`` (pdfplumber words plus exact printed
rectangle topology).  TSN raw and normalized workbooks are streamed with the
generic audit-owned OOXML reader and projected from an independently declared
contract.  Duplicate pairing uses the standard-library-only Phase-3 oracle.

Highway Detail's TSMIS Excel export has no County column.  This audit never
invents one.  An Excel row receives a County only from a uniquely printed
current companion signature, a separately versioned exact historical TSMIS
companion signature over the identical current row payload, or a uniquely
matched component of a current PDF composite Description under one printed DCR
owner.  Every attribution tier is retained separately.  This is source
attestation, not a license to fall back to the product's Route+Post-Mile key or
to use TSN to manufacture the TSMIS side of its own comparison.

This file currently publishes a non-accepting source-comparison draft.  The
five isolated production legs, workbook inspection, permanent mutation gate,
and detached publication decision are attached only after this source truth is
adversarially reviewed and frozen.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_EVEN, localcontext
from difflib import SequenceMatcher
import hashlib
import importlib.metadata
import json
from pathlib import Path
import re
import subprocess
import sys
from typing import Iterable, Sequence
import zipfile
import zlib
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


BUILD_ROOT = Path(__file__).resolve().parent
REPO_ROOT = BUILD_ROOT.parent
GENERATOR_PATH = Path(__file__).resolve()
sys.path.insert(0, str(BUILD_ROOT))

from phase3_independent_oracle import (  # noqa: E402
    FieldRule,
    OracleCounts,
    OracleOutcome,
    OracleRow,
    OracleSchema,
    ValueRule,
    canonical_key,
    compare_row,
    compare_rows as _reference_compare_rows,
    normalize_value,
    pair_group,
)
from phase3_xlsx_stream import (  # noqa: E402
    DATE,
    SCALAR,
    ColumnSpec,
    SheetSpec,
    XlsxLimits,
    read_sheet,
)
import phase8_highway_detail_source_oracle as tsmis_source  # noqa: E402


PRIVATE_ROOT = tsmis_source.PRIVATE_ROOT
DEFAULT_TSMIS_XLSX_ROOT = PRIVATE_ROOT / "tsmis_excel"
DEFAULT_TSMIS_PDF_ROOT = PRIVATE_ROOT / "tsmis_pdf"
DEFAULT_TSN_RAW = (
    PRIVATE_ROOT / "tsn_raw" / "TSAR - HIGHWAY DETAIL_TSN.xlsx")
DEFAULT_TSN_PDF_ROOT = PRIVATE_ROOT / "tsn_pdf"
DEFAULT_ORIGIN_TSMIS_XLSX_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9"
    r"\2026-07-09 ars-prod\highway_detail")
DEFAULT_ORIGIN_TSMIS_PDF_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9"
    r"\2026-07-09 ars-prod\highway_detail_pdf")
DEFAULT_ORIGIN_HISTORICAL_OWNER_XLSX_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth"
    r"\Hwy Detail Dev Bundle 7.7\TSMIS\highway_detail")
DEFAULT_ORIGIN_HISTORICAL_OWNER_PDF_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth"
    r"\Hwy Detail Dev Bundle 7.7\TSMIS\highway_detail_pdf")
DEFAULT_ORIGIN_TSN_RAW = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\highway_detail\raw"
    r"\TSAR - HIGHWAY DETAIL_TSN.xlsx")
DEFAULT_ORIGIN_TSN_PDF_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\highway_detail\pdf")

PHASE4_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd\phase4_tsn_rebaseline"
)
HISTORICAL_OWNER_ROOT = (
    PHASE4_ROOT.parent / "phase8_highway_detail_historical_7_7_owner_sources_r1")
DEFAULT_HISTORICAL_OWNER_XLSX_ROOT = HISTORICAL_OWNER_ROOT / "tsmis_excel"
DEFAULT_HISTORICAL_OWNER_PDF_ROOT = HISTORICAL_OWNER_ROOT / "tsmis_pdf"
DEFAULT_TSN_NORMALIZED = (
    PHASE4_ROOT / "raw-2026-07-12-r7" / "highway_detail" /
    "consolidated" / "tsn_highway_detail_normalized.xlsx")
DEFAULT_TSN_NORMALIZED_SIDECAR = Path(
    str(DEFAULT_TSN_NORMALIZED) + ".outcome.json")
PHASE6_ROOT = PHASE4_ROOT.parent / "phase6_tsn_conservation"
DEFAULT_STAGE6_RESULT = PHASE6_ROOT / "highway_detail_conservation_r7.json"
DEFAULT_STAGE6_ACCEPTANCE = Path(str(DEFAULT_STAGE6_RESULT) + ".acceptance.json")
DEFAULT_TSN_CROSS_FORMAT = PHASE4_ROOT / "highway_detail_tsn_pdf_oracle_final.json"
DEFAULT_OUTPUT = PHASE4_ROOT / "phase8_highway_detail_comparison_source_draft.json"
DEFAULT_PRODUCT_ROOT = (
    REPO_ROOT / "tmp" / "phase8-highway-detail-product-r1")
PRODUCT_HELPER_PATH = BUILD_ROOT / "phase8_highway_detail_product_witness.py"
SELF_GATE_PATH = BUILD_ROOT / "check_phase8_highway_detail_comparison.py"
AUDIT_CODE_PATHS = {
    "generator": GENERATOR_PATH,
    "source_oracle": BUILD_ROOT / "phase8_highway_detail_source_oracle.py",
    "independent_comparator": BUILD_ROOT / "phase3_independent_oracle.py",
    "immutable_xlsx_reader": BUILD_ROOT / "phase3_xlsx_stream.py",
    "product_helper": PRODUCT_HELPER_PATH,
    "self_gate": SELF_GATE_PATH,
}


FILE_BINDINGS = {
    "tsn_raw": {
        "bytes": 16_356_075,
        "sha256": "bac3c882002b26433e39fad00c3dcdf9ad95b8dfc9ba9597386c656a71071dd1",
    },
    "tsn_normalized": {
        "bytes": 8_478_589,
        "sha256": "46afd2b20c08113636eb69630065672afc1044dba02afeab445ac9f0afac34d5",
    },
    "tsn_normalized_sidecar": {
        "bytes": 900,
        "sha256": "97a9ccff48d446eab5d4a16d4383bd7858025fd3022cf4a111cbbe0481175327",
    },
    "stage6_result": {
        "bytes": 122_006,
        "sha256": "283315b30605461e748246444ea523542f61b0a205cd70131c73e1f6b77fb20b",
    },
    "stage6_acceptance": {
        "bytes": 3_802,
        "sha256": "d26dee5d11517478312cde6361c4567c30a4f8d534d822539bb36388c170cf03",
    },
    "tsn_cross_format": {
        "bytes": 664_322,
        "sha256": "540b1ce575be880f506ebc435acaabe253e238f4eba312a72a310129f4ecdc36",
    },
}
TSN_PDF_TREE_BINDING = {
    "files": 12,
    "bytes": 42_667_451,
    "manifest_sha256": (
        "92e91831be1c399af1630a5f9937c2fa2770e203438ad28d51db1ef0df1c3a46"),
}
HISTORICAL_OWNER_TREE_BINDINGS = {
    "xlsx": {
        "files": 2, "bytes": 3_887_485,
        "manifest_sha256": (
            "234b57d9e07c72f1355cb2abc5f91fd952d440172f054d25a81322bf7a284821"),
    },
    "pdf": {
        "files": 2, "bytes": 1_549_484,
        "manifest_sha256": (
            "9ab6b7daca9eb26ade07305781d237739a17fb39da9b6718cc09e0b5c5cc14eb"),
    },
}

_EXPECTED_PDF_TSN_FIELDS = {
    "AC": 166, "Acc-Cont Eff": 1190, "City": 498,
    "Date of Rec": 2847, "Description": 3045, "District": 103,
    "HG": 1601, "LB #Ln": 1950, "LB Eff": 14091,
    "LB IN-TO": 2060, "LB IN-TR": 2062, "LB OT-TO": 4879,
    "LB OT-TR": 4422, "LB S/F": 3830, "LB S/T": 3851,
    "LB Wid": 3637, "Length": 7604, "Med B": 6330,
    "Med C": 3421, "Med Eff": 15719, "Med T": 6357,
    "Med V/WDA": 6507, "NA": 99, "PS": 326,
    "RB #Ln": 5567, "RB Eff": 15840, "RB IN-TO": 4825,
    "RB IN-TR": 4847, "RB OT-TO": 5639, "RB OT-TR": 5601,
    "RB S/F": 4169, "RB S/T": 6705, "RB Wid": 5710,
    "RU": 92, "RU Eff": 47730,
}
_EXPECTED_EXCEL_TSN_FIELDS = {
    "AC": 166, "Acc-Cont Eff": 1188, "City": 498,
    "Date of Rec": 2836, "Description": 3015, "District": 103,
    "HG": 1601, "LB #Ln": 1840, "LB Eff": 13971,
    "LB IN-TO": 2036, "LB IN-TR": 2038, "LB OT-TO": 4856,
    "LB OT-TR": 4400, "LB S/F": 3827, "LB S/T": 3740,
    "LB Wid": 3527, "Length": 7426, "Med B": 6204,
    "Med C": 3427, "Med Eff": 15612, "Med T": 6227,
    "Med V/WDA": 6378, "NA": 98, "PS": 321,
    "RB #Ln": 5521, "RB Eff": 15745, "RB IN-TO": 4745,
    "RB IN-TR": 4767, "RB OT-TO": 5559, "RB OT-TR": 5521,
    "RB S/F": 4171, "RB S/T": 6600, "RB Wid": 5663,
    "RU": 92, "RU Eff": 47694,
}
EXPECTED_SOURCE_COUNTS = {
    "raw_vs_normalized": {
        "known": True, "paired_rows": 60083, "side_a_only_rows": 0,
        "side_b_only_rows": 0, "differing_rows": 1,
        "differing_cells": 1, "per_field_counts": {"Length": 1},
        "asserted_cells": 2162988, "context_cells": 0,
    },
    "pdf_vs_tsn_raw": {
        "known": True, "paired_rows": 48163, "side_a_only_rows": 3053,
        "side_b_only_rows": 11920, "differing_rows": 48010,
        "differing_cells": 203320,
        "per_field_counts": _EXPECTED_PDF_TSN_FIELDS,
        "asserted_cells": 1733868, "context_cells": 0,
    },
    "pdf_vs_tsn_normalized": {
        "known": True, "paired_rows": 48163, "side_a_only_rows": 3053,
        "side_b_only_rows": 11920, "differing_rows": 48010,
        "differing_cells": 203320,
        "per_field_counts": _EXPECTED_PDF_TSN_FIELDS,
        "asserted_cells": 1733868, "context_cells": 0,
    },
    "excel_attested_vs_tsn_raw": {
        "known": True, "paired_rows": 48127, "side_a_only_rows": 2624,
        "side_b_only_rows": 11956, "differing_rows": 47974,
        "differing_cells": 201413,
        "per_field_counts": _EXPECTED_EXCEL_TSN_FIELDS,
        "asserted_cells": 1732572, "context_cells": 0,
    },
    "excel_attested_vs_tsn_normalized": {
        "known": True, "paired_rows": 48127, "side_a_only_rows": 2624,
        "side_b_only_rows": 11956, "differing_rows": 47974,
        "differing_cells": 201413,
        "per_field_counts": _EXPECTED_EXCEL_TSN_FIELDS,
        "asserted_cells": 1732572, "context_cells": 0,
    },
}
EXPECTED_SOURCE_LEDGERS = {
    "raw_vs_normalized": (
        "747ef9615def542f240fbe8811136a053290e7ab09a7a0a0c8c12239c917891b"),
    "pdf_vs_tsn_raw": (
        "a774088344f5de5fe9d18a5075281af98b4a2aedc49bd52311cae36ddabee546"),
    "pdf_vs_tsn_normalized": (
        "35f76a0a63a170efbe62fa61a46d485e46934dcf734cb92c101b7ff6ecbac299"),
    "excel_attested_vs_tsn_raw": (
        "6b4a58fc8bc89745a223caa4e53ed1c0d675f6e57e42bd5a5ae9fc95df78045a"),
    "excel_attested_vs_tsn_normalized": (
        "aa267aa1f671d5e0108973c6302d604f65111d0f945cb547852fb409c234190e"),
}
EXPECTED_TSMIS_SOURCE = {
    "excel_rows": 51273,
    "excel_ordered_typed_rows_sha256": (
        "85162791b8f38adcceb254eafdbf293e9a957a7c9dcfdef9c7e1d8777b562070"),
    "pdf_rows": 51216,
    "pdf_ordered_typed_rows_sha256": (
        "033c5010ef93c8a01df9aa733390bf7af6d158ca26fb195ebf18346e69e1db3b"),
    "pdf_reconciliation": {
        "blocks_deriving_line1_from_line2": 4,
        "blocks_gridless_header_anchored": 8,
        "blocks_gridless_uniquely_projected": 4,
        "dcr_headers": 4065, "furniture_groups": 25902,
        "line1_records": 51216, "line2_physical_groups": 51218,
        "line2_records": 51216, "multigroup_line2_physical_groups": 3,
        "multigroup_line2_records": 1, "pages_with_both_grids": 3664,
        "pages_with_line2_only": 17, "pages_with_neither_grid": 15,
    },
    "format_totals": {
        "all_34_render_equal": 50203, "excel_only_rows": 497,
        "excel_rows": 51273, "fields_2_through_33_render_equal": 573,
        "paired_differing_cells": 701, "paired_differing_rows": 573,
        "paired_rows": 50776, "pdf_only_rows": 440, "pdf_rows": 51216,
    },
    "format_pair_map_sha256": (
        "5b36602cac3255acdc698952fc270975d0f31ea93320ac50a31d0ad9e5bf4192"),
}
EXPECTED_HISTORICAL_OWNER_SOURCE = {
    "excel_rows": 3_174,
    "excel_ordered_typed_rows_sha256": (
        "68cc2ad809dac6326c1218eee45f2913acda968abf8653c58043fbbee902ca19"),
    "pdf_rows": 3_172,
    "pdf_ordered_typed_rows_sha256": (
        "22d03f2a72705bb3e81f2d799f15d8bda940bae6b79edd5a5165f9809a4f898c"),
    "pdf_reconciliation": {
        "dcr_headers": 233, "furniture_groups": 1_561,
        "line1_records": 3_172, "line2_physical_groups": 3_172,
        "line2_records": 3_172, "pages_with_both_grids": 222,
    },
    "current_excel_to_historical_pdf_totals": {
        "all_34_render_equal": 3_167, "excel_only_rows": 5,
        "excel_rows": 3_174, "fields_2_through_33_render_equal": 2,
        "paired_differing_cells": 2, "paired_differing_rows": 2,
        "paired_rows": 3_169, "pdf_only_rows": 3, "pdf_rows": 3_172,
    },
    "current_excel_to_historical_pdf_pair_map_sha256": (
        "d9249c8a0e8abd5958fb804a12cd7887cbb9baad10c5046ba18c088a3870dc95"),
    "historical_excel_to_pdf_pair_map_sha256": (
        "dea2b5e1cb2a1a240f68f6e3d0e7a0edea3f1bc56001dd9dbad3b4c10d9de48d"),
    "same_build_route_005_owner_rows": 3_125,
    "route_005_current_and_historical_excel_sha256": (
        "00a359555c964f46a68f36b32ae1a44501168eaee553911aa85838b9afef24c5"),
}
EXPECTED_PRODUCT_OBJECT_SHA256 = (
    "abd1fcbab499f38cb923e4f68d4da89f62c5e5889feae6705b06f478a013a340")
EXPECTED_OWNER_CONSTRAINT_SHA256 = (
    "e134db90de11371cbff95507c549d6969ec6423b9f7a42896d28f214c03d1045")
EXPECTED_OWNER_COMPARISON_CONTRACT_SHA256 = {
    "snapshot_attested_excel_vs_tsn_raw": (
        "bd190bb0a45195a296bfc4a45ff88fa066539492f6edaf364b60d9cccdd91bdf"),
    "snapshot_attested_excel_vs_tsn_normalized": (
        "2e9b08dc81b5d528ef1f879ae66d8b2cfc9981cbeb6a608aecb3bda9b4f7dea4"),
}


RAW_HEADERS = (
    "THY_ID", "DIST", "CNTY", "RTE", "RTE_SFX", "DIST_CNTY_ROUTE",
    "PP", "POSTMILE", "E_IND", "LENGTH", "REC_DATE", "HG", "AC",
    "ACC_SIG", "ACC_EFF_DATE", "CITY", "POP_CODE", "BEG_DATE",
    "ADT_AMT", "PROFILE", "BREAK_DESC", "LK_BACK_ADT", "CHNGMILE",
    "DVM", "DESCRIPTION", "NON_ADD", "LT_SIG", "L_EFF_DATE", "L_ST",
    "L_NO_LANES", "L_SF", "L_OT_TOT", "L_OT_TR", "L_TR_WID",
    "L_IN_TOT", "L_IN_TR", "MED_SIG", "M_EFF_DATE", "M_TYPE_CODE",
    "M_CL", "M_BA", "M_WID", "M_VA", "RT_SIG", "R_EFF_DATE",
    "R_ST", "R_NO_LANES", "R_SF", "R_IN_TOT", "R_IN_TR", "R_TR_WID",
    "R_OT_TOT", "R_OT_TR", "SEG_ORDER_ID", "REFERENCE_DATE",
    "EXTRACT_DATE",
)
NORMALIZED_HEADERS = (
    "Route", "Post Mile", "PS", "Length", "Date of Rec", "HG", "AC",
    "Acc-Cont Eff", "City", "RU", "RU Eff", "Description", "NA",
    "LB Eff", "LB S/T", "LB #Ln", "LB S/F", "LB OT-TO", "LB OT-TR",
    "LB Wid", "LB IN-TO", "LB IN-TR", "Med Eff", "Med T", "Med C",
    "Med B", "Med V/WDA", "RB Eff", "RB S/T", "RB #Ln", "RB S/F",
    "RB IN-TO", "RB IN-TR", "RB Wid", "RB OT-TO", "RB OT-TR",
    "TSN District", "TSN County",
)
RAW_SPEC = SheetSpec(
    "Sheet 1",
    tuple(ColumnSpec(
        header,
        DATE if header in {"REFERENCE_DATE", "EXTRACT_DATE"} else SCALAR,
    ) for header in RAW_HEADERS),
    exact_schema=True,
)
NORMALIZED_SPEC = SheetSpec(
    "Highway Detail (TSN)",
    tuple(ColumnSpec(header, SCALAR) for header in NORMALIZED_HEADERS),
    exact_schema=True,
)

TSMIS_HEADERS = tsmis_source.TSMIS_HEADERS
SHARED_FIELDS = ("PS", *TSMIS_HEADERS[1:])
SOURCE_ASSERTED_FIELDS = ("District", "County", *SHARED_FIELDS)
SOURCE_SCHEMA = OracleSchema(
    key_rules=(
        ValueRule("Route"), ValueRule("County"), ValueRule("Complete PP"),
        ValueRule("Numeric Post Mile"), ValueRule("Roadbed"),
    ),
    field_rules=tuple(FieldRule(field, asserting=True)
                      for field in SOURCE_ASSERTED_FIELDS),
)
PRODUCT_SCHEMA = OracleSchema(
    key_rules=(ValueRule("Route"), ValueRule("Post Mile")),
    field_rules=tuple(FieldRule(field, asserting=True)
                      for field in SHARED_FIELDS),
)
CONSOLIDATED_SPEC = SheetSpec(
    "Highway Detail",
    tuple(ColumnSpec(header, SCALAR)
          for header in ("Route", *TSMIS_HEADERS)),
    exact_schema=True,
)

RAW_COLUMN = {
    "Length": "LENGTH", "Date of Rec": "REC_DATE", "HG": "HG",
    "AC": "AC", "Acc-Cont Eff": "ACC_EFF_DATE", "City": "CITY",
    "RU": "POP_CODE", "RU Eff": "BEG_DATE", "Description": "DESCRIPTION",
    "NA": "NON_ADD", "LB Eff": "L_EFF_DATE", "LB S/T": "L_ST",
    "LB #Ln": "L_NO_LANES", "LB S/F": "L_SF", "LB OT-TO": "L_OT_TOT",
    "LB OT-TR": "L_OT_TR", "LB Wid": "L_TR_WID",
    "LB IN-TO": "L_IN_TOT", "LB IN-TR": "L_IN_TR",
    "Med Eff": "M_EFF_DATE", "Med T": "M_TYPE_CODE", "Med C": "M_CL",
    "Med B": "M_BA", "RB Eff": "R_EFF_DATE", "RB S/T": "R_ST",
    "RB #Ln": "R_NO_LANES", "RB S/F": "R_SF", "RB IN-TO": "R_IN_TOT",
    "RB IN-TR": "R_IN_TR", "RB Wid": "R_TR_WID",
    "RB OT-TO": "R_OT_TOT", "RB OT-TR": "R_OT_TR",
}
DATE_FIELDS = {
    "Date of Rec", "Acc-Cont Eff", "RU Eff", "LB Eff", "Med Eff", "RB Eff",
}
NUMERIC_FIELDS = {
    "LB #Ln", "LB OT-TO", "LB OT-TR", "LB Wid", "LB IN-TO", "LB IN-TR",
    "RB #Ln", "RB OT-TO", "RB OT-TR", "RB Wid", "RB IN-TO", "RB IN-TR",
}
SOURCE_ONLY_FIELDS = (
    "THY_ID", "DIST_CNTY_ROUTE", "ACC_SIG", "ADT_AMT", "PROFILE",
    "BREAK_DESC", "LK_BACK_ADT", "CHNGMILE", "DVM", "LT_SIG", "MED_SIG",
    "RT_SIG", "SEG_ORDER_ID", "REFERENCE_DATE", "EXTRACT_DATE",
)

PM_TOKEN_RE = re.compile(
    r"^([A-Z]{0,2})(\d{1,3}\.\d{3})([A-Z]{0,2})$", re.IGNORECASE)
ROUTE_RE = re.compile(r"^(\d{1,3})([A-Z]?)$", re.IGNORECASE)


class AuditError(RuntimeError):
    """A bound source or independently declared audit contract drifted."""


@dataclass(frozen=True)
class DetailRow:
    source_index: int
    source: str
    source_ref: str
    route: str
    district: str
    county: str
    complete_pp: str
    numeric_pm: str
    equation: str
    roadbed: str
    explicit_trailing: str
    values: tuple[object, ...]
    source_only: tuple[tuple[str, object], ...] = ()
    raw_values: tuple[object, ...] = ()

    @property
    def physical_key(self) -> tuple[str, str, str, str, str]:
        if not self.county:
            raise AuditError(
                f"county-unknown row cannot form physical identity: {self.source_ref}")
        return (
            self.route, self.county, self.complete_pp,
            self.numeric_pm, self.roadbed,
        )

    @property
    def canonical_pm(self) -> str:
        return f"{self.complete_pp}{self.numeric_pm}{self.roadbed}"

    def oracle_row(self) -> OracleRow:
        return OracleRow(
            source_index=self.source_index,
            key=self.physical_key,
            values=(self.district, self.county, *self.values),
            source_ref=self.source_ref,
        )

    def product_oracle_row(self) -> OracleRow:
        return OracleRow(
            source_index=self.source_index,
            key=(self.route, self.canonical_pm),
            values=self.values,
            source_ref=self.source_ref,
        )

    def product_projection(self) -> tuple[object, ...]:
        return (self.route, self.canonical_pm, *self.values)


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
        default=str,
    ).encode("utf-8")


def _text(value: object) -> str:
    if value is None:
        return ""
    if type(value) is bool:
        return "TRUE" if value else "FALSE"
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise AuditError("non-finite Decimal source value")
        return format(value, "f")
    return str(value).strip()


def _route(value: object, suffix: object = "") -> str:
    token = (_text(value) + _text(suffix)).upper()
    match = ROUTE_RE.fullmatch(token)
    if match is None:
        raise AuditError(f"invalid Highway Detail route: {token!r}")
    return f"{int(match.group(1)):03d}{match.group(2).upper()}"


def _district(value: object) -> str:
    token = _text(value)
    if token.isdigit():
        return f"{int(token):02d}"
    if token in {"\u2014", "\u2013", "-"}:
        return token
    raise AuditError(f"invalid Highway Detail district: {token!r}")


def _fixed_three(value: object, label: str) -> str:
    literal = _text(value)
    try:
        number = Decimal(literal)
    except InvalidOperation as exc:
        raise AuditError(f"invalid {label} decimal: {literal!r}") from exc
    if not number.is_finite():
        raise AuditError(f"non-finite {label} decimal: {literal!r}")
    with localcontext() as context:
        context.prec = max(50, len(number.as_tuple().digits) + 10)
        rounded = number.quantize(Decimal("0.001"), rounding=ROUND_HALF_EVEN)
    return format(rounded, "07.3f")


def _pm_parts(value: object, hg: object) -> tuple[str, str, str, str]:
    token = _text(value).upper()
    match = PM_TOKEN_RE.fullmatch(token)
    if match is None:
        raise AuditError(f"invalid glued Highway Detail Post Mile: {token!r}")
    pp, numeric, trailing = match.groups()
    equation = "E" if "E" in trailing else ""
    explicit = next((char for char in reversed(trailing)
                     if char in {"R", "L"}), "")
    hg_roadbed = _text(hg).upper()
    roadbed = explicit or (hg_roadbed if hg_roadbed in {"R", "L"} else "")
    return pp.upper(), _fixed_three(numeric, "Post Mile"), equation, roadbed


def _date_value(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%y-%m-%d")
    return _text(value)


def _numeric_value(value: object) -> str:
    token = _text(value)
    return str(int(token)) if token.isdigit() else token


def _wda_value(value: object) -> str:
    token = _text(value).upper()
    match = re.fullmatch(r"(\d+)([A-Z]?)", token)
    return (f"{int(match.group(1)):02d}{match.group(2)}"
            if match is not None else token)


def _project(field: str, value: object) -> str:
    if field in DATE_FIELDS:
        return _date_value(value)
    if field in NUMERIC_FIELDS:
        return _numeric_value(value)
    if field == "Length":
        token = _text(value)
        if not token:
            return ""
        try:
            return _fixed_three(value, "Length")
        except AuditError:
            # Vendor-pending TSMIS currently includes literal non-decimal print
            # shapes.  Preserve them as claims; never coerce them to a number.
            return token
    if field == "NA":
        token = _text(value).upper()
        return "" if token == "A" else token
    if field == "Med V/WDA":
        return _wda_value(value)
    if field == "Description":
        return " ".join(_text(value).replace("\u00a0", " ").split())
    return _text(value)


def _from_tsmis_source(
        row: tsmis_source.SourceRow, *,
        district: str | None = None, county: str | None = None) -> DetailRow:
    values = tuple(row.values)
    if len(values) != len(TSMIS_HEADERS):
        raise AuditError(f"TSMIS row width drift at {row.source_ref}")
    raw = dict(zip(TSMIS_HEADERS, values))
    pp, numeric_pm, equation, roadbed = _pm_parts(
        raw["Post Mile"], raw["HG"])
    trailing = PM_TOKEN_RE.fullmatch(_text(raw["Post Mile"]).upper()).group(3)
    projected = (equation, *(
        _project(field, raw[field]) for field in TSMIS_HEADERS[1:]))
    return DetailRow(
        source_index=row.source_index,
        source=row.source,
        source_ref=row.source_ref,
        route=_route(row.member_route),
        district=(district if district is not None else row.district),
        county=(county if county is not None else row.county).upper().rstrip("."),
        complete_pp=pp,
        numeric_pm=numeric_pm,
        equation=equation,
        roadbed=roadbed,
        explicit_trailing=trailing,
        values=tuple(projected),
        raw_values=values,
    )


def _from_raw(values: Sequence[object], source_index: int,
              source_row: int) -> DetailRow:
    if len(values) != len(RAW_HEADERS):
        raise AuditError(f"raw TSN row width drift at worksheet row {source_row}")
    raw = dict(zip(RAW_HEADERS, values))
    route = _route(raw["RTE"], raw["RTE_SFX"])
    pp = _text(raw["PP"]).upper()
    numeric_pm = _fixed_three(raw["POSTMILE"], "raw TSN Post Mile")
    equation = _text(raw["E_IND"]).upper()
    if equation not in {"", "E"}:
        raise AuditError(f"unknown raw TSN E_IND {equation!r} at row {source_row}")
    hg = _text(raw["HG"]).upper()
    roadbed = hg if hg in {"R", "L"} else ""
    projected: list[object] = []
    for field in SHARED_FIELDS:
        if field == "PS":
            projected.append(equation)
        elif field == "Med V/WDA":
            projected.append(_wda_value(
                _text(raw["M_WID"]) + _text(raw["M_VA"])))
        else:
            projected.append(_project(field, raw[RAW_COLUMN[field]]))
    return DetailRow(
        source_index=source_index,
        source="TSN raw XLSX",
        source_ref=(f"{DEFAULT_TSN_RAW.name}:row {source_row}"),
        route=route,
        district=_district(raw["DIST"]),
        county=_text(raw["CNTY"]).upper().rstrip("."),
        complete_pp=pp,
        numeric_pm=numeric_pm,
        equation=equation,
        roadbed=roadbed,
        explicit_trailing=equation,
        values=tuple(projected),
        source_only=tuple((field, raw[field]) for field in SOURCE_ONLY_FIELDS),
        raw_values=tuple(values),
    )


def _from_normalized(values: Sequence[object], source_index: int,
                     source_row: int) -> DetailRow:
    if len(values) != len(NORMALIZED_HEADERS):
        raise AuditError(
            f"normalized TSN row width drift at worksheet row {source_row}")
    raw = dict(zip(NORMALIZED_HEADERS, values))
    route = _route(raw["Route"])
    pp, numeric_pm, token_equation, roadbed = _pm_parts(
        raw["Post Mile"], raw["HG"])
    equation = _text(raw["PS"]).upper()
    if equation not in {"", "E"}:
        raise AuditError(
            f"unknown normalized PS {equation!r} at row {source_row}")
    if token_equation and token_equation != equation:
        raise AuditError(
            f"normalized Post Mile/PS conflict at row {source_row}")
    projected = tuple(
        equation if field == "PS" else _project(field, raw[field])
        for field in SHARED_FIELDS)
    trailing = PM_TOKEN_RE.fullmatch(_text(raw["Post Mile"]).upper()).group(3)
    return DetailRow(
        source_index=source_index,
        source="TSN normalized r7",
        source_ref=f"{DEFAULT_TSN_NORMALIZED.name}:row {source_row}",
        route=route,
        district=_district(raw["TSN District"]),
        county=_text(raw["TSN County"]).upper().rstrip("."),
        complete_pp=pp,
        numeric_pm=numeric_pm,
        equation=equation,
        roadbed=roadbed,
        explicit_trailing=trailing,
        values=projected,
        raw_values=tuple(values),
    )


def _parse_tsn_raw(path: Path) -> tuple[list[DetailRow], dict[str, object]]:
    sheet = read_sheet(
        path, RAW_SPEC, limits=XlsxLimits(max_xml_events=25_000_000))
    rows = [
        _from_raw(physical.values, index, physical.source_row)
        for index, physical in enumerate(sheet.rows)
    ]
    if len(rows) != 60_083:
        raise AuditError(f"raw TSN row count drift: {len(rows)}")
    return rows, _source_summary("TSN raw XLSX", rows, {
        "sheet": sheet.sheet_name,
        "columns": len(sheet.headers),
        "source_rows_contiguous_from_2": (
            [row.source_ref.rsplit(" ", 1)[-1] for row in rows[:1]] == ["2"]
            and sheet.rows[-1].source_row == len(rows) + 1),
    })


def _parse_tsn_normalized(path: Path) -> tuple[list[DetailRow], dict[str, object]]:
    sheet = read_sheet(
        path, NORMALIZED_SPEC, limits=XlsxLimits(max_xml_events=25_000_000))
    rows = [
        _from_normalized(physical.values, index, physical.source_row)
        for index, physical in enumerate(sheet.rows)
    ]
    if len(rows) != 60_083:
        raise AuditError(f"normalized TSN row count drift: {len(rows)}")
    return rows, _source_summary("TSN normalized r7", rows, {
        "sheet": sheet.sheet_name,
        "columns": len(sheet.headers),
        "source_rows_contiguous_from_2": sheet.rows[-1].source_row == len(rows) + 1,
    })


def _row_payload(row: DetailRow) -> tuple[object, ...]:
    return (
        row.route, row.district, row.county, row.complete_pp, row.numeric_pm,
        row.equation, row.roadbed, row.explicit_trailing, *row.values,
        *row.source_only, *row.raw_values,
    )


def _rows_digest(rows: Iterable[DetailRow]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update(_canonical(_row_payload(row)))
        digest.update(b"\n")
    return digest.hexdigest()


def _multiplicity(counter: Counter[tuple[object, ...]]) -> dict[str, int]:
    duplicates = [count for count in counter.values() if count > 1]
    return {
        "unique": len(counter),
        "duplicate_groups": len(duplicates),
        "duplicate_occurrences": sum(duplicates),
        "occurrences_beyond_first": sum(count - 1 for count in duplicates),
        "max_multiplicity": max(counter.values(), default=0),
    }


def _identity_census(rows: Sequence[DetailRow]) -> dict[str, object]:
    known = [row for row in rows if row.county]
    strong = Counter(row.physical_key for row in known)
    equation = Counter((*row.physical_key, row.equation) for row in known)
    weak_counties: defaultdict[tuple[str, str], set[str]] = defaultdict(set)
    base_pp_counties: defaultdict[tuple[str, str, str], set[str]] = defaultdict(set)
    for row in known:
        weak_counties[(row.route, row.canonical_pm)].add(row.county)
        base_pp_counties[(row.route[:3], row.complete_pp, row.numeric_pm)].add(
            row.county)
    weak = [value for value in weak_counties.values() if len(value) > 1]
    base = [value for value in base_pp_counties.values() if len(value) > 1]
    return {
        "county_known_rows": len(known),
        "county_unknown_rows": len(rows) - len(known),
        "physical_without_equation": _multiplicity(strong),
        "physical_with_equation": _multiplicity(equation),
        "route_canonical_pm_cross_county": {
            "keys": len(weak),
            "county_identities": sum(len(item) for item in weak),
            "max_counties": max((len(item) for item in weak), default=0),
        },
        "base_route_pp_pm_cross_county": {
            "keys": len(base),
            "county_identities": sum(len(item) for item in base),
            "max_counties": max((len(item) for item in base), default=0),
        },
        "district_census": dict(sorted(Counter(row.district for row in known).items())),
        "county_census": dict(sorted(Counter(row.county for row in known).items())),
        "route_census": dict(sorted(Counter(row.route for row in rows).items())),
        "equation_census": dict(sorted(Counter(row.equation for row in rows).items())),
        "roadbed_census": dict(sorted(Counter(row.roadbed for row in rows).items())),
        "explicit_trailing_census": dict(sorted(Counter(
            row.explicit_trailing for row in rows).items())),
    }


def _source_summary(label: str, rows: Sequence[DetailRow],
                    extra: dict[str, object]) -> dict[str, object]:
    return {
        "label": label,
        "rows": len(rows),
        "ordered_source_payload_sha256": _rows_digest(rows),
        "identity": _identity_census(rows),
        "null_or_empty_by_asserted_field": {
            field: sum(_text(value) == "" for value in values)
            for field, values in zip(
                SHARED_FIELDS,
                zip(*(row.values for row in rows), strict=True),
            )
        } if rows else {},
        **extra,
    }


def _render_cell(value: object) -> str:
    return " ".join(_text(value).replace("\u00a0", " ").split())


def _source_signature(row: tsmis_source.SourceRow,
                      excluded: frozenset[int]) -> tuple[str, ...]:
    return tuple(_render_cell(value) for index, value in enumerate(row.values)
                 if index not in excluded)


def _format_pairs(
        excel_rows: Sequence[tsmis_source.SourceRow],
        pdf_rows: Sequence[tsmis_source.SourceRow],
        public_alignment: dict[str, object]) -> tuple[
            list[tuple[tsmis_source.SourceRow, tsmis_source.SourceRow, str]],
            dict[str, object]]:
    excel_by_member: defaultdict[str, list[tsmis_source.SourceRow]] = defaultdict(list)
    pdf_by_member: defaultdict[str, list[tsmis_source.SourceRow]] = defaultdict(list)
    for row in excel_rows:
        excel_by_member[row.member_route].append(row)
    for row in pdf_rows:
        pdf_by_member[row.member_route].append(row)
    if set(excel_by_member) != set(pdf_by_member):
        raise AuditError("TSMIS format member universes differ")

    found: list[tuple[tsmis_source.SourceRow, tsmis_source.SourceRow, str]] = []
    digest = hashlib.sha256()
    for member in sorted(excel_by_member):
        left = excel_by_member[member]
        right = pdf_by_member[member]
        left_full = [_source_signature(row, frozenset()) for row in left]
        right_full = [_source_signature(row, frozenset()) for row in right]
        first = SequenceMatcher(
            None, left_full, right_full, autojunk=False).get_opcodes()
        pairs: dict[int, tuple[int, str]] = {}
        paired_right: set[int] = set()
        changed = []
        for tag, i1, i2, j1, j2 in first:
            if tag == "equal":
                for left_index, right_index in zip(range(i1, i2), range(j1, j2)):
                    pairs[left_index] = (right_index, "all_34_render_equal")
                    paired_right.add(right_index)
            else:
                changed.append((i1, i2, j1, j2))
        for i1, i2, j1, j2 in changed:
            left_claims = [_source_signature(row, frozenset({0, 1}))
                           for row in left[i1:i2]]
            right_claims = [_source_signature(row, frozenset({0, 1}))
                            for row in right[j1:j2]]
            second = SequenceMatcher(
                None, left_claims, right_claims, autojunk=False).get_opcodes()
            for tag, a1, a2, b1, b2 in second:
                if tag != "equal":
                    continue
                for left_offset, right_offset in zip(range(a1, a2), range(b1, b2)):
                    left_index, right_index = i1 + left_offset, j1 + right_offset
                    if left_index in pairs or right_index in paired_right:
                        raise AuditError("non-bijective TSMIS format pair map")
                    pairs[left_index] = (
                        right_index, "fields_2_through_33_render_equal")
                    paired_right.add(right_index)
        ordered = sorted(
            (left_index, right_index, classification)
            for left_index, (right_index, classification) in pairs.items())
        if any(b2 <= b1 for (_a1, b1, _c1), (_a2, b2, _c2)
               in zip(ordered, ordered[1:])):
            raise AuditError(f"{member}: non-monotonic TSMIS format map")
        for left_index, right_index, classification in ordered:
            left_row, right_row = left[left_index], right[right_index]
            found.append((left_row, right_row, classification))
            digest.update(_canonical((
                member, left_index, right_index, classification,
                right_row.district, right_row.county,
            )))
            digest.update(b"\n")
    observed_digest = digest.hexdigest()
    expected_digest = public_alignment["pair_map_sha256"]
    if observed_digest != expected_digest:
        raise AuditError(
            f"independent TSMIS pair map drift: {observed_digest} != {expected_digest}")
    if len(found) != public_alignment["totals"]["paired_rows"]:
        raise AuditError("independent/public TSMIS pair count drift")
    return found, {
        "pairs": len(found),
        "pair_map_sha256": observed_digest,
        "matches_public_source_oracle": True,
    }


def _attest_excel_county(
        excel_rows: Sequence[tsmis_source.SourceRow],
        pdf_rows: Sequence[tsmis_source.SourceRow],
        pairs: Sequence[tuple[
            tsmis_source.SourceRow, tsmis_source.SourceRow, str]],
        ) -> tuple[list[DetailRow], dict[str, object]]:
    owners_by_signature: defaultdict[
        tuple[str, str, tuple[str, ...]], set[tuple[str, str]]
    ] = defaultdict(set)
    for row in pdf_rows:
        owners_by_signature[(
            row.member_route, "all_34_render_equal",
            _source_signature(row, frozenset()),
        )].add((row.district, row.county))
        owners_by_signature[(
            row.member_route, "fields_2_through_33_render_equal",
            _source_signature(row, frozenset({0, 1})),
        )].add((row.district, row.county))

    attested: list[DetailRow] = []
    ambiguous = []
    paired_excel = set()
    classifications = Counter()
    for excel, pdf, classification in pairs:
        paired_excel.add(excel.source_index)
        excluded = (frozenset() if classification == "all_34_render_equal"
                    else frozenset({0, 1}))
        owners = owners_by_signature[(
            excel.member_route, classification,
            _source_signature(excel, excluded),
        )]
        if owners != {(pdf.district, pdf.county)}:
            if len(ambiguous) < 100:
                ambiguous.append({
                    "excel_ref": excel.source_ref,
                    "pdf_ref": pdf.source_ref,
                    "classification": classification,
                    "candidate_printed_owners": [list(owner)
                                                  for owner in sorted(owners)],
                })
            continue
        attested.append(_from_tsmis_source(
            excel, district=pdf.district, county=pdf.county))
        classifications[classification] += 1

    unpaired = [row for row in excel_rows if row.source_index not in paired_excel]
    return attested, {
        "rule": (
            "an Excel row receives district/county only when its exact paired "
            "companion-PDF signature has exactly one printed DCR owner within "
            "that route member"),
        "all_format_pairs": len(pairs),
        "uniquely_owner_attested_rows": len(attested),
        "attested_by_classification": dict(classifications),
        "ambiguous_cross_owner_pair_count": len(pairs) - len(attested),
        "ambiguous_cross_owner_samples_first_100": ambiguous,
        "format_unpaired_excel_rows": len(unpaired),
        "format_unpaired_excel_samples_first_100": [{
            "source_ref": row.source_ref,
            "member": row.member_route,
            "postmile": row.postmile,
        } for row in unpaired[:100]],
        "total_county_unknown_excel_rows": (
            len(excel_rows) - len(attested)),
        "county_inference_used": False,
    }


def _composite_description_owner_attestations(
        excel_rows: Sequence[tsmis_source.SourceRow],
        pdf_rows: Sequence[tsmis_source.SourceRow],
        eligible_source_indices: set[int],
        ) -> tuple[
            dict[int, tuple[str, str, dict[str, object]]],
            dict[str, object]]:
    """Attest only exact unique Excel components of a printed PDF composite.

    A slash-delimited PDF Description is still one PDF record.  Matching one
    component does not manufacture another PDF row or copy any non-owner cell.
    It proves only that the exact current Excel Description is visibly printed
    beneath that parent row's DCR owner.
    """
    excel_by_description: defaultdict[
        tuple[str, str], list[tsmis_source.SourceRow]
    ] = defaultdict(list)
    for row in excel_rows:
        description = _render_cell(row.values[9])
        if description:
            excel_by_description[(row.member_route, description)].append(row)

    claims: defaultdict[
        int, list[tuple[str, str, dict[str, object]]]
    ] = defaultdict(list)
    composite_rows = total_components = unique_excel_components = 0
    for pdf in pdf_rows:
        description = _render_cell(pdf.values[9])
        components = tuple(
            component.strip() for component in description.split(" / ")
            if component.strip())
        if len(components) < 2:
            continue
        composite_rows += 1
        total_components += len(components)
        for ordinal, component in enumerate(components):
            candidates = excel_by_description[(pdf.member_route, component)]
            if len(candidates) != 1:
                continue
            unique_excel_components += 1
            candidate = candidates[0]
            if candidate.source_index not in eligible_source_indices:
                continue
            claims[candidate.source_index].append((
                pdf.district, pdf.county, {
                    "excel_source_ref": candidate.source_ref,
                    "excel_postmile": candidate.postmile,
                    "component": component,
                    "component_ordinal": ordinal,
                    "parent_pdf_source_ref": pdf.source_ref,
                    "parent_pdf_postmile": pdf.postmile,
                    "parent_pdf_description": description,
                    "parent_pdf_district": pdf.district,
                    "parent_pdf_county": pdf.county,
                    "parent_component_count": len(components),
                }))

    chosen: dict[int, tuple[str, str, dict[str, object]]] = {}
    ambiguous = []
    digest = hashlib.sha256()
    for source_index in sorted(claims):
        candidates = claims[source_index]
        if (len(candidates) == 1 and candidates[0][0]
                and candidates[0][1]):
            chosen[source_index] = candidates[0]
            digest.update(_canonical((source_index, candidates[0])))
            digest.update(b"\n")
        else:
            ambiguous.append({
                "source_index": source_index,
                "claims": [list(item[:2]) for item in candidates],
                "parent_refs": [item[2]["parent_pdf_source_ref"]
                                for item in candidates],
            })
    return chosen, {
        "rule": (
            "split only current PDF Descriptions containing a literal spaced "
            "slash separator; match a nonempty component to exactly one "
            "current Excel Description in the same route member; promote only "
            "one eligible claim beneath one nonblank printed DCR owner; attest "
            "owner only, never synthesize a PDF record or other cells"),
        "composite_pdf_rows": composite_rows,
        "components_in_composite_rows": total_components,
        "components_unique_in_current_excel_member": unique_excel_components,
        "eligible_component_claim_rows": len(claims),
        "uniquely_owner_attested_rows": len(chosen),
        "ambiguous_eligible_rows": len(ambiguous),
        "ambiguous": ambiguous,
        "attestations": [chosen[index][2] for index in sorted(chosen)],
        "attestation_ledger_sha256": digest.hexdigest(),
        "row_equivalence_claimed": False,
    }


def _analyze_excel_owner_constraints(
        excel_rows: Sequence[tsmis_source.SourceRow],
        pdf_rows: Sequence[tsmis_source.SourceRow],
        raw_rows: Sequence[DetailRow],
        exactly_attested: Sequence[DetailRow],
        historical_exactly_attested: Sequence[DetailRow],
        ) -> tuple[list[DetailRow], dict[str, object]]:
    """Classify every County-less Excel row without silently inventing owner.

    Exact companion-signature attestation remains the strongest tier.  For the
    residue, a row is separately marked *companion-key constrained* only when
    every current TSMIS PDF row with the same observable physical components
    has one identical printed District/County owner.  TSN-only candidates are
    reported but never promoted: using the comparison target to manufacture a
    missing TSMIS identity would be circular.
    """
    excel_detail = [
        _from_tsmis_source(row, district="", county="")
        for row in excel_rows]
    excel_detail_by_index = {row.source_index: row for row in excel_detail}
    pdf_detail = [_from_tsmis_source(row) for row in pdf_rows]

    def ownerless(row: DetailRow) -> tuple[str, str, str, str]:
        return (row.route, row.complete_pp, row.numeric_pm, row.roadbed)

    pdf_owners: defaultdict[
        tuple[str, str, str, str], set[tuple[str, str]]
    ] = defaultdict(set)
    tsn_owners: defaultdict[
        tuple[str, str, str, str], set[tuple[str, str]]
    ] = defaultdict(set)
    for row in pdf_detail:
        pdf_owners[ownerless(row)].add((row.district, row.county))
    for row in raw_rows:
        tsn_owners[ownerless(row)].add((row.district, row.county))

    exact = {row.source_index: row for row in exactly_attested}
    historical = {
        row.source_index: row for row in historical_exactly_attested}
    overlap = set(exact).intersection(historical)
    historical_conflicts = [{
        "source_index": source_index,
        "source_ref": excel_detail_by_index[source_index].source_ref,
        "current_later_pdf_owner": [
            exact[source_index].district, exact[source_index].county],
        "same_build_historical_pdf_owner": [
            historical[source_index].district,
            historical[source_index].county],
    } for source_index in sorted(overlap)
        if (exact[source_index].district, exact[source_index].county)
        != (historical[source_index].district,
            historical[source_index].county)]
    historical_conflict_digest = hashlib.sha256()
    for item in historical_conflicts:
        historical_conflict_digest.update(_canonical(item))
        historical_conflict_digest.update(b"\n")

    preliminary: dict[int, dict[str, object]] = {}
    for row in excel_detail:
        key = ownerless(row)
        printed = pdf_owners.get(key, set())
        tsn = tsn_owners.get(key, set())
        chosen: tuple[str, str] | None = None
        if row.source_index in historical:
            classification = (
                "same_build_historical_exact_companion_unique_owner")
            chosen = (historical[row.source_index].district,
                      historical[row.source_index].county)
        elif row.source_index in exact:
            classification = "exact_companion_signature_unique_owner"
            chosen = (exact[row.source_index].district,
                      exact[row.source_index].county)
        elif len(printed) == 1:
            classification = "companion_key_single_owner"
            chosen = next(iter(printed))
        elif len(printed) > 1:
            classification = "companion_key_multiple_owners"
        elif len(tsn) == 1:
            classification = "no_companion_key_tsn_single_owner_not_promoted"
        elif len(tsn) > 1:
            classification = "no_companion_key_tsn_multiple_owners"
        else:
            classification = "no_owner_candidate"
        preliminary[row.source_index] = {
            "classification": classification,
            "chosen": chosen,
            "key": key,
            "printed": printed,
            "tsn": tsn,
        }

    eligible = {
        source_index for source_index, item in preliminary.items()
        if item["chosen"] is None}
    composite, composite_summary = (
        _composite_description_owner_attestations(
            excel_rows, pdf_rows, eligible))

    extended: list[DetailRow] = []
    classifications: Counter[str] = Counter()
    per_route: defaultdict[str, Counter[str]] = defaultdict(Counter)
    agreement: Counter[str] = Counter()
    samples = []
    digest = hashlib.sha256()
    for row in excel_detail:
        item = preliminary[row.source_index]
        classification = str(item["classification"])
        chosen = item["chosen"]
        key = item["key"]
        printed = item["printed"]
        tsn = item["tsn"]
        composite_evidence = None
        if chosen is None and row.source_index in composite:
            district, county, composite_evidence = composite[row.source_index]
            chosen = (district, county)
            classification = (
                "current_pdf_composite_description_unique_owner")
        if printed and tsn:
            agreement[
                "same_owner_set" if printed == tsn else
                "overlapping_owner_sets" if printed.intersection(tsn) else
                "disjoint_owner_sets"] += 1
        elif printed:
            agreement["companion_only"] += 1
        elif tsn:
            agreement["tsn_only"] += 1
        else:
            agreement["neither"] += 1
        classifications[classification] += 1
        per_route[row.route][classification] += 1
        if chosen is not None:
            extended.append(replace(
                row, district=chosen[0], county=chosen[1]))
        if classification not in {
                "exact_companion_signature_unique_owner"} and len(samples) < 100:
            samples.append({
                "source_ref": row.source_ref,
                "observable_key": list(key),
                "classification": classification,
                "companion_pdf_owners": [list(owner)
                                          for owner in sorted(printed)],
                "tsn_owners": [list(owner) for owner in sorted(tsn)],
                "promoted_owner": list(chosen) if chosen else None,
                "composite_evidence": composite_evidence,
            })
        digest.update(_canonical((
            row.source_index, key, classification, sorted(printed),
            sorted(tsn), chosen, composite_evidence)))
        digest.update(b"\n")
    unresolved = len(excel_detail) - len(extended)
    return extended, {
        "rule": (
            "promote an exact uniquely owned current companion signature; "
            "then an exact separately versioned historical TSMIS companion "
            "signature over the current Excel payload; then a unanimous "
            "current-companion-PDF owner at the observable key; finally one "
            "unique exact component of a current composite PDF Description; "
            "retain TSN-only candidates as non-promoted diagnostics"),
        "excel_rows": len(excel_detail),
        "classifications": dict(classifications),
        "source_owner_set_relation": dict(agreement),
        "owner_attested_rows": len(extended),
        "new_companion_key_constrained_rows": classifications[
            "companion_key_single_owner"],
        "same_build_historical_owner_rows": len(historical),
        "current_exact_rows_superseded_by_same_build_snapshot": len(overlap),
        "same_build_historical_owner_rows_promoted": classifications[
            "same_build_historical_exact_companion_unique_owner"],
        "cross_edition_owner_conflicts": len(historical_conflicts),
        "cross_edition_owner_conflict_ledger_sha256":
            historical_conflict_digest.hexdigest(),
        "cross_edition_owner_conflicts_exact": historical_conflicts,
        "composite_description_attestation": composite_summary,
        "unresolved_owner_rows": unresolved,
        "classification_ledger_sha256": digest.hexdigest(),
        "per_route": {
            route: dict(counts) for route, counts in sorted(per_route.items())},
        "non_exact_samples_first_100": samples,
        "tsn_only_owner_promotions": 0,
        "county_inference_from_filename_or_order": False,
    }


def _compare_rows_indexed(
        schema: OracleSchema, rows_a: Sequence[OracleRow],
        rows_b: Sequence[OracleRow], *, pair_cap: int = 100_000,
        ) -> OracleOutcome:
    """Semantic twin of Phase-3 ``compare_rows`` with O(1) order membership.

    CMP-AUD-187 owns the shared implementation's quadratic ``key not in list``
    membership.  The list still records exact first-seen order here; a set is
    used only to decide whether that key has already been appended.
    """
    groups_a: dict[tuple[object, ...], list[OracleRow]] = {}
    groups_b: dict[tuple[object, ...], list[OracleRow]] = {}
    order: list[tuple[object, ...]] = []
    seen: set[tuple[object, ...]] = set()
    for rows, groups in ((rows_a, groups_a), (rows_b, groups_b)):
        for row in rows:
            key = canonical_key(row, schema)
            if key not in seen:
                seen.add(key)
                order.append(key)
            groups.setdefault(key, []).append(row)

    paired_rows = side_a_only = side_b_only = 0
    differing_rows = differing_cells = 0
    asserted_cells = context_cells = 0
    per_field = {rule.name: 0 for rule in schema.field_rules if rule.asserting}
    row_results = []
    side_a_indices = []
    side_b_indices = []
    traces = []
    diagnostics = []
    asserting_width = sum(rule.asserting for rule in schema.field_rules)
    context_width = len(schema.field_rules) - asserting_width

    for key in order:
        left = groups_a.get(key, [])
        right = groups_b.get(key, [])
        if not left:
            side_b_only += len(right)
            side_b_indices.extend(row.source_index for row in right)
            continue
        if not right:
            side_a_only += len(left)
            side_a_indices.extend(row.source_index for row in left)
            continue
        paired = pair_group(left, right, schema, key, pair_cap=pair_cap)
        traces.append(paired.trace)
        if paired.capped_diagnostic is not None:
            diagnostics.append(paired.capped_diagnostic)
        for left_position, right_position in paired.pairs:
            result = compare_row(
                left[left_position], right[right_position], schema, key)
            row_results.append(result)
            paired_rows += 1
            asserted_cells += asserting_width
            context_cells += context_width
            differing_rows += bool(result.differing_fields)
            for field in result.differing_fields:
                per_field[field] += 1
                differing_cells += 1
        side_a_only += len(paired.unmatched_a)
        side_b_only += len(paired.unmatched_b)
        side_a_indices.extend(left[index].source_index
                              for index in paired.unmatched_a)
        side_b_indices.extend(right[index].source_index
                              for index in paired.unmatched_b)

    no_data = not rows_a and not rows_b
    completion = "no_data" if no_data else "partial" if diagnostics else "complete"
    verdict = (
        "unknown" if no_data else
        "match" if completion == "complete" and not differing_cells
        and not side_a_only and not side_b_only else "diff")
    counts = OracleCounts(
        known=True,
        paired_rows=paired_rows,
        side_a_only_rows=side_a_only,
        side_b_only_rows=side_b_only,
        differing_rows=differing_rows,
        differing_cells=differing_cells,
        per_field_counts={name: count for name, count in per_field.items() if count},
        asserted_cells=asserted_cells,
        context_cells=context_cells,
    )
    return OracleOutcome(
        completion=completion,
        verdict=verdict,
        counts=counts,
        row_results=tuple(row_results),
        side_a_only_indices=tuple(side_a_indices),
        side_b_only_indices=tuple(side_b_indices),
        pairing_trace=tuple(traces),
        pairing_quality="capped" if diagnostics else "exact",
        capped_diagnostics=tuple(diagnostics),
    )


def _indexed_oracle_equivalence_gate() -> dict[str, object]:
    schema = OracleSchema(
        key_rules=(ValueRule("K1"), ValueRule("K2")),
        field_rules=(FieldRule("A"), FieldRule("B")),
    )

    def row(index: int, key: tuple[object, object],
            values: tuple[object, object]) -> OracleRow:
        return OracleRow(index, key, values, f"gate:{index}")

    cases = [
        ("empty", (), (), 100_000),
        ("unique_and_one_sided",
         (row(0, ("001", "A"), ("x", "1")),
          row(1, ("002", "B"), ("y", "2"))),
         (row(10, ("001", "A"), ("x", "9")),
          row(11, ("003", "C"), ("z", "3"))), 100_000),
        ("duplicate_minimum_cost",
         (row(0, ("005", "X"), ("alpha", "1")),
          row(1, ("005", "X"), ("beta", "2"))),
         (row(10, ("005", "X"), ("beta", "2")),
          row(11, ("005", "X"), ("alpha", "1"))), 100_000),
        ("typed_and_reordered_keys",
         (row(0, (Decimal("1.0"), " Q "), (None, "a  b")),
          row(1, (2, "R"), ("", "c"))),
         (row(10, (2, "R"), (None, "c")),
          row(11, (1, "Q"), ("", "a b"))), 100_000),
        ("capped_duplicate_assignment",
         (row(0, ("009", "Z"), ("a", "1")),
          row(1, ("009", "Z"), ("b", "2"))),
         (row(10, ("009", "Z"), ("a", "2")),
          row(11, ("009", "Z"), ("b", "1"))), 1),
    ]
    evidence = []
    for label, left, right, cap in cases:
        expected = _reference_compare_rows(schema, left, right, pair_cap=cap)
        actual = _compare_rows_indexed(schema, left, right, pair_cap=cap)
        exact = asdict(expected) == asdict(actual)
        evidence.append({
            "case": label,
            "exact": exact,
            "outcome_sha256": _sha_bytes(_canonical(asdict(actual))),
        })
    if not all(item["exact"] for item in evidence):
        raise AuditError(f"indexed oracle equivalence failed: {evidence!r}")
    return {
        "cases": evidence,
        "passed": len(evidence),
        "required": len(evidence),
        "exact": True,
    }


def _comparison(label: str, left: Sequence[DetailRow],
                right: Sequence[DetailRow]) -> dict[str, object]:
    outcome = _compare_rows_indexed(
        SOURCE_SCHEMA,
        tuple(row.oracle_row() for row in left),
        tuple(row.oracle_row() for row in right),
    )
    by_left = {row.source_index: row for row in left}
    by_right = {row.source_index: row for row in right}
    digest = hashlib.sha256()
    examples = []
    examples_by_field: defaultdict[str, list[dict[str, object]]] = defaultdict(list)
    for result in outcome.row_results:
        left_row = by_left[result.source_index_a]
        right_row = by_right[result.source_index_b]
        entry = {
            "key": list(left_row.physical_key),
            "left_ref": left_row.source_ref,
            "right_ref": right_row.source_ref,
            "differing_fields": list(result.differing_fields),
            "differences": [{
                "field": field,
                "left": cell.normalized_a.text,
                "right": cell.normalized_b.text,
            } for field, cell in zip(SOURCE_ASSERTED_FIELDS, result.cells)
                if cell.counts_as_difference],
        }
        digest.update(_canonical(entry))
        digest.update(b"\n")
        if result.differing_fields and len(examples) < 100:
            examples.append(entry)
        for difference in entry["differences"]:
            field = str(difference["field"])
            if len(examples_by_field[field]) < 10:
                examples_by_field[field].append({
                    "key": entry["key"],
                    "left_ref": entry["left_ref"],
                    "right_ref": entry["right_ref"],
                    **difference,
                })
    left_only = [by_left[index] for index in outcome.side_a_only_indices]
    right_only = [by_right[index] for index in outcome.side_b_only_indices]
    return {
        "label": label,
        "completion": outcome.completion,
        "verdict": outcome.verdict,
        "counts": asdict(outcome.counts),
        "pairing_quality": outcome.pairing_quality,
        "duplicate_pair_groups": sum(
            trace.side_a_size > 1 or trace.side_b_size > 1
            for trace in outcome.pairing_trace),
        "max_pair_matrix_cells": max(
            (trace.matrix_cells for trace in outcome.pairing_trace), default=0),
        "capped_diagnostics": [asdict(item)
                               for item in outcome.capped_diagnostics],
        "ordered_pair_ledger_sha256": digest.hexdigest(),
        "difference_examples_first_100": examples,
        "difference_examples_first_10_by_field": dict(sorted(
            examples_by_field.items())),
        "left_only_samples_first_100": [{
            "key": list(row.physical_key), "source_ref": row.source_ref,
        } for row in left_only[:100]],
        "right_only_samples_first_100": [{
            "key": list(row.physical_key), "source_ref": row.source_ref,
        } for row in right_only[:100]],
    }


def _product_text(value: object) -> str:
    normalized = normalize_value(value)
    return "" if normalized.kind == "blank" else normalized.text


def _counter_digest(counter: Counter[tuple[object, ...]]) -> str:
    ordered = sorted(counter.items(), key=lambda item: _canonical(item[0]))
    return _sha_bytes(_canonical([[key, count] for key, count in ordered]))


def _counter_difference_examples(
        observed: Counter[tuple[object, ...]],
        expected: Counter[tuple[object, ...]]) -> dict[str, object]:
    missing = expected - observed
    extra = observed - expected

    def first(counter: Counter[tuple[object, ...]]) -> list[dict[str, object]]:
        return [
            {"entry": list(key), "count": count}
            for key, count in sorted(
                counter.items(), key=lambda item: _canonical(item[0]))[:10]
        ]

    return {"missing_first_10": first(missing), "extra_first_10": first(extra)}


def _product_expected(
        label: str, side_a: str, side_b: str,
        left: Sequence[DetailRow], right: Sequence[DetailRow]
        ) -> dict[str, object]:
    """Derive the exact weak-key result the current product is required to emit.

    This deliberately models the product's declared Route+Post-Mile identity,
    not the source-authoritative county-backed identity.  The two ledgers remain
    separate so a self-consistent weak-key workbook cannot certify source truth.
    """
    outcome = _compare_rows_indexed(
        PRODUCT_SCHEMA,
        tuple(row.product_oracle_row() for row in left),
        tuple(row.product_oracle_row() for row in right),
    )
    if outcome.completion != "complete" or outcome.pairing_quality != "exact":
        raise AuditError(f"{label}: weak product oracle did not complete exactly")
    counts = asdict(outcome.counts)
    paired: Counter[tuple[object, ...]] = Counter()
    for result in outcome.row_results:
        displays = []
        state = []
        for cell in result.cells:
            if cell.equal:
                displays.append(
                    "" if cell.normalized_a.kind == "blank"
                    else cell.normalized_a.text)
            else:
                displays.append(cell.display)
            state.append("E" if cell.equal else "D")
        key = tuple(
            "" if item.kind == "blank" else item.text for item in result.key)
        paired[(*key, len(result.differing_fields), *displays,
                "".join(state))] += 1

    by_left = {row.source_index: row for row in left}
    by_right = {row.source_index: row for row in right}

    def one_sided(indices: Sequence[int], rows: dict[int, DetailRow]
                  ) -> Counter[tuple[object, ...]]:
        result: Counter[tuple[object, ...]] = Counter()
        for index in indices:
            row = rows[index]
            result[(
                _product_text(row.route), _product_text(row.canonical_pm),
                *(_product_text(value) for value in row.values),
            )] += 1
        return result

    side_a_only = one_sided(outcome.side_a_only_indices, by_left)
    side_b_only = one_sided(outcome.side_b_only_indices, by_right)
    pairing_trace = json.loads(_canonical([
        asdict(trace) for trace in outcome.pairing_trace
        if trace.side_a_size > 1 or trace.side_b_size > 1]))
    pairing_semantics = [{
        "algorithm": trace["algorithm"],
        "quality": trace["quality"],
        "exact": trace["exact"],
        "key_components": [
            "" if item["kind"] == "blank" else item["text"]
            for item in trace["key"]],
        "matrix_cells": trace["matrix_cells"],
        "side_a_size": trace["side_a_size"],
        "side_b_size": trace["side_b_size"],
        "smaller_side": trace["smaller_side"],
        "total_cost": trace["total_cost"],
        "source_pairs": sorted(trace["source_pairs"]),
    } for trace in pairing_trace]
    return {
        "label": label, "side_a": side_a, "side_b": side_b,
        "counts": counts,
        "pairing": {
            "quality": outcome.pairing_quality,
            "groups": len(outcome.pairing_trace),
            "duplicate_groups": sum(
                trace.side_a_size > 1 or trace.side_b_size > 1
                for trace in outcome.pairing_trace),
            "max_matrix_cells": max(
                (trace.matrix_cells for trace in outcome.pairing_trace),
                default=0),
        },
        "pairing_trace_sha256": _sha_bytes(_canonical(pairing_semantics)),
        "independent_pairing_trace_wire_sha256": _sha_bytes(
            _canonical(pairing_trace)),
        "paired_ledger_sha256": _counter_digest(paired),
        "side_a_only_ledger_sha256": _counter_digest(side_a_only),
        "side_b_only_ledger_sha256": _counter_digest(side_b_only),
        "_paired": paired,
        "_side_a_only": side_a_only,
        "_side_b_only": side_b_only,
        "_pairing_trace": pairing_trace,
        "_pairing_trace_semantics": pairing_semantics,
    }


def _public_product_expected(expected: dict[str, object]) -> dict[str, object]:
    return {key: value for key, value in expected.items()
            if not key.startswith("_")}


def _parse_product_consolidated(
        path: Path, source_label: str) -> tuple[list[DetailRow], object]:
    sheet = read_sheet(
        path, CONSOLIDATED_SPEC,
        limits=XlsxLimits(max_xml_events=25_000_000))
    rows: list[DetailRow] = []
    for physical in sheet.rows:
        values = tuple(physical.values)
        source = tsmis_source.SourceRow(
            source_index=len(rows), source=source_label,
            source_ref=f"{path.name}:row {physical.source_row}",
            member_route=_route(values[0]), district="", county="",
            values=values[1:],
        )
        rows.append(_from_tsmis_source(source, district="", county=""))
    return rows, sheet


def _raw_cell(value: object) -> tuple[str, str]:
    if value in (None, ""):
        return "blank", ""
    if isinstance(value, datetime):
        return "datetime", value.isoformat()
    if isinstance(value, date):
        return "date", value.isoformat()
    if isinstance(value, Decimal):
        return "Decimal", format(value, "f")
    return type(value).__name__, str(value)


def _raw_row(values: Sequence[object]) -> tuple[tuple[str, str], ...]:
    return tuple(_raw_cell(value) for value in values)


def _inspect_consolidated_against_source(
        path: Path, truth_rows: Sequence[DetailRow], source_label: str
        ) -> tuple[dict[str, object], list[DetailRow]]:
    actual_rows, sheet = _parse_product_consolidated(path, source_label)
    expected_raw = [
        _raw_row((row.route, *row.raw_values)) for row in truth_rows]
    observed_raw = [
        _raw_row((row.route, *row.raw_values)) for row in actual_rows]
    expected_inventory = Counter(expected_raw)
    observed_inventory = Counter(observed_raw)
    source_only_inventory = expected_inventory - observed_inventory
    product_only_inventory = observed_inventory - expected_inventory

    def routes(counter: Counter[tuple[tuple[str, str], ...]]) -> dict[str, int]:
        result: Counter[str] = Counter()
        for row, count in counter.items():
            result[row[0][1]] += count
        return dict(sorted(result.items()))

    first_ordered_mismatch = None
    ordered_mismatches = abs(len(expected_raw) - len(observed_raw))
    for offset, (expected, observed) in enumerate(
            zip(expected_raw, observed_raw), 2):
        if expected != observed:
            ordered_mismatches += 1
            if first_ordered_mismatch is None:
                differing = [index for index, pair in enumerate(
                    zip(expected, observed), 1) if pair[0] != pair[1]]
                first_ordered_mismatch = {
                    "worksheet_row": offset,
                    "differing_columns": differing,
                    "expected": [list(item) for item in expected],
                    "observed": [list(item) for item in observed],
                }
    semantic = _product_expected(
        f"{source_label} source vs product consolidation",
        "Source", "Product consolidation", truth_rows, actual_rows)
    return ({
        "path": str(path.resolve()),
        "sheet": sheet.sheet_name,
        "columns": len(sheet.headers),
        "source_rows": len(truth_rows),
        "product_rows": len(actual_rows),
        "ordered_raw_rows_exact": expected_raw == observed_raw,
        "ordered_raw_row_mismatches": ordered_mismatches,
        "first_ordered_raw_mismatch": first_ordered_mismatch,
        "raw_multiset_exact": expected_inventory == observed_inventory,
        "raw_multiset_difference_examples": _counter_difference_examples(
            observed_inventory, expected_inventory),
        "raw_multiset_difference_counts": {
            "source_only_rows": sum(source_only_inventory.values()),
            "product_only_rows": sum(product_only_inventory.values()),
            "source_only_by_route": routes(source_only_inventory),
            "product_only_by_route": routes(product_only_inventory),
        },
        "source_only_raw_ledger_sha256": _counter_digest(
            source_only_inventory),
        "product_only_raw_ledger_sha256": _counter_digest(
            product_only_inventory),
        "expected_raw_multiset_sha256": _counter_digest(expected_inventory),
        "observed_raw_multiset_sha256": _counter_digest(observed_inventory),
        "independent_projected_semantic_reconciliation":
            _public_product_expected(semantic),
        "source_projection_exact": (
            semantic["counts"]["side_a_only_rows"] == 0
            and semantic["counts"]["side_b_only_rows"] == 0
            and semantic["counts"]["differing_cells"] == 0),
    }, actual_rows)


def _formula_tag_census(path: Path) -> dict[str, int]:
    pattern = re.compile(rb"<(?:[A-Za-z0-9_]+:)?f(?:\s|>)")
    counts: dict[str, int] = {}
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(
            archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships}
        ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        for sheet in workbook.findall("main:sheets/main:sheet", ns):
            label = sheet.attrib["name"]
            relation = sheet.attrib[f"{{{ns['rel']}}}id"]
            target = targets[relation].replace("\\", "/")
            member_name = (target.lstrip("/") if target.startswith("/")
                           else "xl/" + target)
            count = 0
            tail = b""
            with archive.open(member_name) as member:
                for chunk in iter(lambda: member.read(1024 * 1024), b""):
                    payload = tail + chunk
                    count += sum(
                        match.end() > len(tail)
                        for match in pattern.finditer(payload))
                    tail = payload[-96:]
            counts[label] = count
    return counts


def _formula_flavor_contract(
        formula_census: dict[str, int], value_census: dict[str, int],
        expected_sheets: Sequence[str], expected_formula: dict[str, int],
        expected_values: dict[str, int], side_a: str, side_b: str,
        ) -> bool:
    """Exact, non-contradictory formula/value workbook census contract."""
    return (
        list(formula_census) == list(expected_sheets)
        and list(value_census) == list(expected_sheets)
        and formula_census == expected_formula
        and value_census == expected_values
        and sum(formula_census.values()) > sum(value_census.values()) > 0
        and formula_census["Comparison"] > value_census["Comparison"]
        and formula_census[side_a] > value_census[side_a]
        and formula_census[side_b] > value_census[side_b])


def _validate_file_identity(
        identity: dict[str, object], expected_path: Path, label: str
        ) -> dict[str, object]:
    actual = {
        "path": str(expected_path.resolve()),
        "bytes": expected_path.stat().st_size,
        "sha256": _sha_file(expected_path),
    }
    declared_path = Path(str(identity.get("path", ""))).resolve()
    if (declared_path != expected_path.resolve()
            or identity.get("bytes") != actual["bytes"]
            or identity.get("sha256") != actual["sha256"]):
        raise AuditError(
            f"{label}: product-declared artifact identity is stale: "
            f"{identity!r} vs {actual!r}")
    return actual


PRODUCT_HEADER = ("Post Mile", *SHARED_FIELDS)
MEDWID_HELPER_HEADERS = tuple(
    f"__CMP_E1_MW_V1_F025_{stage}"
    for stage in ("TRIM", "CORE", "VALID", "MASK", "CANON"))
PAYLOAD_BASENAME_RE = re.compile(
    r"\.cmpv3-[0-9a-f]{64}-[0-9]{6}-[0-9a-f]{64}"
    r"\.comparison-payload\.zlib")


def _excel_trim(value: object) -> str:
    if value is None:
        return ""
    if type(value) is bool:
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(" +", " ", str(value)).strip(" ")


def _medwid_canonical(value: str) -> str:
    match = re.fullmatch(r"([0-9]+)(?:\.([0-9]+))?(.)?", value)
    if match is None:
        return value
    whole, fraction, suffix = match.groups()
    if suffix is not None and not (
            0x21 <= ord(suffix) < 0x7F
            and suffix not in "0123456789."):
        return value
    whole = whole.lstrip("0") or "0"
    if fraction is not None:
        fraction = fraction.rstrip("0")
    number = whole + (f".{fraction}" if fraction else "")
    return number + (suffix or "")


def _medwid_helper_values(value: object) -> tuple[object, ...]:
    trimmed = _excel_trim(value)
    has_suffix = bool(trimmed) and (
        0x21 <= ord(trimmed[-1]) < 0x7F
        and trimmed[-1] not in "0123456789.")
    core = trimmed[:-1] if has_suffix else trimmed
    valid = re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", core) is not None
    mask = ("".join("X" if char == "." or char in "123456789" else "0"
                    for char in core) if valid else "")
    return trimmed, core, valid, mask, _medwid_canonical(trimmed)


def _projection_tuple(row: DetailRow) -> tuple[str, ...]:
    return tuple(_product_text(value) for value in row.product_projection())


def _inspect_projection_sheet(
        worksheet, rows: Sequence[DetailRow], *, side: str,
        snapshot: bool) -> dict[str, object]:
    physical = worksheet.iter_rows(values_only=True)
    header = tuple(next(physical, ()))
    expected_header = (
        ("Source row", "Route", *PRODUCT_HEADER, "Key (helper)",
         *MEDWID_HELPER_HEADERS)
        if snapshot else
        ("Comparison row", "Route", *PRODUCT_HEADER, "Key (helper)",
         *MEDWID_HELPER_HEADERS,
         "__CMP_E2_BUILD_FRESH_V1_C001_B_AQ"))
    if header != expected_header:
        raise AuditError(f"{worksheet.title}: projected header drift: {header!r}")
    digest = hashlib.sha256()
    expected_digest = hashlib.sha256()
    count = 0
    helper_tokens = set()
    for physical_row in physical:
        if not physical_row or all(value is None for value in physical_row):
            continue
        if len(physical_row) != len(expected_header):
            raise AuditError(f"{worksheet.title}: projected row width drift")
        if count >= len(rows):
            raise AuditError(f"{worksheet.title}: unexpected extra projected row")
        if snapshot and physical_row[0] != count + 1:
            raise AuditError(
                f"{worksheet.title}: source ordinal drift at {count + 1}")
        projection_start = 1
        projection_stop = projection_start + len(PRODUCT_HEADER) + 1
        helper_index = projection_stop
        medwid_start = helper_index + 1
        medwid_stop = medwid_start + len(MEDWID_HELPER_HEADERS)
        observed = tuple(_product_text(value) for value in
                         physical_row[projection_start:projection_stop])
        expected = _projection_tuple(rows[count])
        if observed != expected:
            raise AuditError(
                f"{worksheet.title}: projected source drift at row {count + 2}: "
                f"{observed!r} != {expected!r}")
        helper = str(physical_row[helper_index] or "")
        if not helper.startswith("__CMP_E2_KEY_V1_") or helper in helper_tokens:
            raise AuditError(f"{worksheet.title}: helper identity drift")
        helper_tokens.add(helper)
        observed_medwid = tuple(
            _product_text(value)
            for value in physical_row[medwid_start:medwid_stop])
        expected_medwid = tuple(
            _product_text(value) for value in _medwid_helper_values(
                rows[count].values[SHARED_FIELDS.index("Med V/WDA")]))
        if observed_medwid != expected_medwid:
            raise AuditError(
                f"{worksheet.title}: Med V/WDA helper drift at row "
                f"{count + 2}: {observed_medwid!r} != {expected_medwid!r}")
        digest.update(_canonical(observed))
        digest.update(b"\n")
        expected_digest.update(_canonical(expected))
        expected_digest.update(b"\n")
        count += 1
    if count != len(rows):
        raise AuditError(
            f"{worksheet.title}: projected rows {count} != {len(rows)}")
    return {
        "side": side, "rows": count, "columns": len(expected_header),
        "ordered_projection_sha256": digest.hexdigest(),
        "expected_ordered_projection_sha256": expected_digest.hexdigest(),
        "projection_exact": True, "helper_tokens_unique": True,
        "medwid_helpers_exact": True,
    }


def _report_view_source_counter(
        rows: Sequence[DetailRow], *, raw: bool
        ) -> Counter[tuple[object, ...]]:
    result: Counter[tuple[object, ...]] = Counter()
    for row in rows:
        values = ["", "", "", "", "", ""]
        if raw:
            claims = dict(row.source_only)
            raw_row = dict(zip(RAW_HEADERS, row.raw_values))
            dcr = _text(claims.get("DIST_CNTY_ROUTE")) or "-".join(
                item for item in (
                    _text(raw_row.get("DIST")), _text(raw_row.get("CNTY")),
                    _text(raw_row.get("RTE")) + _text(raw_row.get("RTE_SFX")),
                ) if item)
            values = [
                dcr, _text(raw_row.get("ADT_AMT")),
                _text(raw_row.get("PROFILE")),
                _text(raw_row.get("LK_BACK_ADT")),
                _text(raw_row.get("CHNGMILE")), _text(raw_row.get("DVM")),
            ]
        result[(
            _product_text(row.route), _product_text(row.canonical_pm),
            *(_product_text(value) for value in values),
        )] += 1
    return result


def _inspect_report_view(
        worksheet, *, expected_records: int, side_a: str,
        tsn_rows: Sequence[DetailRow], raw_source_claims: bool
        ) -> dict[str, object]:
    physical = worksheet.iter_rows(values_only=True)
    headers = [tuple(next(physical, ())) for _ in range(4)]
    if any(len(row) != 28 for row in headers):
        raise AuditError("Highway Detail Report View four-row header width drift")
    header_contract = (
        headers[0][13] == "TSN only"
        and headers[1][13] == "DCR"
        and headers[0][14] == "ADT (TSN only)"
        and tuple(headers[1][14:19])
        == ("LK-AHD", "P", "LK-BACK", "CHG/MILE", "DVM"))
    if not header_contract:
        raise AuditError("Highway Detail Report View source-only header drift")
    rows = [tuple(row) for row in physical
            if row and any(value is not None for value in row)]
    if len(rows) % 2:
        raise AuditError("Highway Detail Report View has an orphan physical row")
    records = len(rows) // 2
    if records != expected_records:
        raise AuditError(
            f"Highway Detail Report View records {records} != "
            f"expected {expected_records}")
    observed: Counter[tuple[object, ...]] = Counter()
    nonblank = Counter()
    tsn_bearing = 0
    for index in range(0, len(rows), 2):
        line_one, line_two = rows[index:index + 2]
        if len(line_one) != 28 or len(line_two) != 28:
            raise AuditError("Highway Detail Report View data width drift")
        if _product_text(line_one[2]) != _product_text(line_two[2]):
            raise AuditError("Highway Detail Report View route identity drift")
        is_side_a_only = (
            _product_text(line_one[0]) == side_a
            and _product_text(line_one[1]) == "only")
        source_claims = tuple(_product_text(value)
                              for value in line_one[13:19])
        for field, value in zip(
                ("DCR", "ADT_AMT", "PROFILE", "LK_BACK_ADT",
                 "CHNGMILE", "DVM"), source_claims):
            nonblank[field] += int(bool(value))
        if not is_side_a_only:
            tsn_bearing += 1
            observed[(
                _product_text(line_one[2]), _product_text(line_one[3]),
                *source_claims,
            )] += 1
        elif any(source_claims):
            raise AuditError(
                "TSMIS-only Report View row contains invented TSN source claims")
    expected = _report_view_source_counter(
        tsn_rows, raw=raw_source_claims)
    if observed != expected:
        raise AuditError(
            "Highway Detail Report View TSN-source mapping drift: "
            f"{_counter_difference_examples(observed, expected)!r}")
    if tsn_bearing != len(tsn_rows):
        raise AuditError(
            f"Highway Detail Report View TSN-bearing rows {tsn_bearing} != "
            f"{len(tsn_rows)}")
    if not raw_source_claims and any(nonblank.values()):
        raise AuditError(
            "normalized Highway Detail Report View recovered omitted source claims")
    return {
        "records": records,
        "physical_data_rows": len(rows),
        "tsn_bearing_records": tsn_bearing,
        "source_only_header_mapping_exact": True,
        "source_only_nonblank_counts": dict(nonblank),
        "source_only_ledger_sha256": _counter_digest(observed),
        "expected_source_only_ledger_sha256": _counter_digest(expected),
        "source_only_mapping_exact": True,
        "raw_source_claims_expected": raw_source_claims,
    }


def _helper_outcome_exact(
        label: str, payload: dict[str, object],
        expected: dict[str, object]) -> dict[str, object]:
    result = payload.get("result")
    if not isinstance(result, dict):
        raise AuditError(f"product {label} outcome is missing")
    counts = result.get("counts")
    if not isinstance(counts, dict):
        raise AuditError(f"product {label} returned no structured counts")
    product_counts = dict(counts)
    product_counts["per_field_counts"] = {
        str(key).split(":", 1)[-1]: value
        for key, value in (counts.get("per_field_counts") or {}).items()
        if value}
    expected_counts = dict(expected["counts"])
    if product_counts != expected_counts:
        raise AuditError(
            f"product {label} returned-count drift: {product_counts!r} != "
            f"{expected_counts!r}")
    expected_verdict = (
        "match" if not expected_counts["differing_cells"]
        and not expected_counts["side_a_only_rows"]
        and not expected_counts["side_b_only_rows"] else "diff")
    if (result.get("status"), result.get("completion"), result.get("verdict"),
            result.get("skipped_inputs"), result.get("failed_inputs")) != (
                "ok", "complete", expected_verdict, 0, 0):
        raise AuditError(f"product {label} outcome state drift: {result!r}")
    if result.get("warnings") or result.get("failures"):
        raise AuditError(f"product {label} reported warnings/failures")
    generation = result.get("artifact_generation")
    if not isinstance(generation, dict) or (
            generation.get("completion"), generation.get("publication_state"),
            generation.get("requested_mode")) != (
                "complete", "committed", "both"):
        raise AuditError(f"product {label} twin generation state drift")
    members = generation.get("members")
    flavors = sorted(
        str(member.get("flavor")) for member in members
        if isinstance(member, dict)) if isinstance(members, list) else []
    if flavors != ["formulas", "values"]:
        raise AuditError(f"product {label} twin manifest drift: {flavors!r}")
    return {
        "status": result.get("status"),
        "completion": result.get("completion"),
        "verdict": result.get("verdict"),
        "counts": product_counts,
        "artifact_generation": {
            "completion": generation.get("completion"),
            "publication_state": generation.get("publication_state"),
            "requested_mode": generation.get("requested_mode"),
            "flavors": flavors,
        },
        "matches_independent_expected_counts": True,
    }


def _publication_member_exact(
        member: object, expected_path: Path, expected_flavor: str,
        label: str, bound_identity: dict[str, object]) -> dict[str, object]:
    if not isinstance(member, dict):
        raise AuditError(f"{label}: publication member is not an object")
    stat = expected_path.stat()
    actual = {
        "path": str(expected_path.resolve()),
        "relative_path": expected_path.name,
        "size": stat.st_size,
        "sha256": bound_identity.get("sha256"),
        "mtime_ns": stat.st_mtime_ns,
        "flavor": expected_flavor,
    }
    if (bound_identity.get("path") != actual["path"]
            or bound_identity.get("bytes") != actual["size"]
            or not isinstance(actual["sha256"], str)):
        raise AuditError(f"{label}: pre-bound workbook identity drift")
    declared_path = Path(str(member.get("path", ""))).resolve()
    if (declared_path != expected_path.resolve()
            or member.get("relative_path") != actual["relative_path"]
            or member.get("size") != actual["size"]
            or member.get("sha256") != actual["sha256"]
            or member.get("mtime_ns") != actual["mtime_ns"]
            or member.get("flavor") != expected_flavor
            or member.get("canonical_path_at_write")
            != str(expected_path.resolve()).casefold()):
        raise AuditError(
            f"{label}: publication member identity drift: {member!r} vs "
            f"{actual!r}")
    expected_role = "canonical" if expected_flavor == "values" else "best_effort"
    if member.get("commit_role") != expected_role:
        raise AuditError(
            f"{label}: publication commit role drift: "
            f"{member.get('commit_role')!r} != {expected_role!r}")
    return {**actual, "commit_role": expected_role,
            "canonical_path_at_write_exact": True}


def _normalized_payload_counts(value: object, label: str) -> dict[str, object]:
    if not isinstance(value, dict):
        raise AuditError(f"{label}: persisted payload counts are absent")
    result = dict(value)
    per_field = value.get("per_field_counts")
    if not isinstance(per_field, dict):
        raise AuditError(f"{label}: persisted per-field counts are absent")
    result["per_field_counts"] = {
        str(key).split(":", 1)[-1]: count
        for key, count in per_field.items() if count}
    return result


def _persisted_trace_semantics(
        value: object, label: str) -> list[dict[str, object]]:
    if not isinstance(value, list):
        raise AuditError(f"{label}: persisted pairing trace is not an array")
    expected_fields = {
        "key_components", "side_a_size", "side_b_size", "matrix_cells",
        "side_a_indices", "side_b_indices", "smaller_side",
        "assignment_vector", "pairs", "total_cost", "positional_cost",
        "algorithm", "exact", "quality",
    }
    result = []
    for ordinal, trace in enumerate(value):
        prefix = f"{label}: pairing trace {ordinal}"
        if not isinstance(trace, dict) or set(trace) != expected_fields:
            raise AuditError(f"{prefix} field universe drift")
        key = trace["key_components"]
        side_a = trace["side_a_indices"]
        side_b = trace["side_b_indices"]
        vector = trace["assignment_vector"]
        pairs = trace["pairs"]
        if (not isinstance(key, list)
                or not all(isinstance(item, str) for item in key)
                or not isinstance(side_a, list)
                or not isinstance(side_b, list)
                or not isinstance(vector, list)
                or not isinstance(pairs, list)):
            raise AuditError(f"{prefix} array shape drift")
        if (not all(type(item) is int and item >= 0 for item in side_a)
                or not all(type(item) is int and item >= 0 for item in side_b)
                or len(set(side_a)) != len(side_a)
                or len(set(side_b)) != len(side_b)):
            raise AuditError(f"{prefix} source-index census drift")
        size_a = trace["side_a_size"]
        size_b = trace["side_b_size"]
        matrix = trace["matrix_cells"]
        total = trace["total_cost"]
        positional = trace["positional_cost"]
        if (type(size_a) is not int or type(size_b) is not int
                or type(matrix) is not int or type(total) is not int
                or type(positional) is not int or size_a <= 0 or size_b <= 0
                or total < 0 or positional < 0
                or size_a != len(side_a) or size_b != len(side_b)
                or matrix != size_a * size_b
                or max(size_a, size_b) <= 1):
            raise AuditError(f"{prefix} rectangular census drift")
        expected_smaller = "a" if size_a <= size_b else "b"
        smaller_size = min(size_a, size_b)
        larger_size = max(size_a, size_b)
        if (trace["smaller_side"] != expected_smaller
                or len(vector) != smaller_size
                or not all(type(item) is int and 0 <= item < larger_size
                           for item in vector)
                or len(set(vector)) != len(vector)
                or len(pairs) != smaller_size):
            raise AuditError(f"{prefix} assignment-vector drift")
        reconstructed = []
        for smaller_index, larger_index in enumerate(vector):
            if expected_smaller == "a":
                reconstructed.append(
                    (side_a[smaller_index], side_b[larger_index]))
            else:
                reconstructed.append(
                    (side_a[larger_index], side_b[smaller_index]))
        observed_pairs = []
        costs = []
        for pair in pairs:
            if (not isinstance(pair, dict)
                    or set(pair) != {"side_a_index", "side_b_index", "cost"}
                    or type(pair["side_a_index"]) is not int
                    or type(pair["side_b_index"]) is not int
                    or type(pair["cost"]) is not int or pair["cost"] < 0):
                raise AuditError(f"{prefix} pair record drift")
            observed_pairs.append(
                (pair["side_a_index"], pair["side_b_index"]))
            costs.append(pair["cost"])
        if (observed_pairs != reconstructed or len(set(observed_pairs))
                != len(observed_pairs) or sum(costs) != total
                or total > positional):
            raise AuditError(f"{prefix} assignment reconstruction drift")
        if (trace["algorithm"] != "rectangular-hungarian-lex-v1"
                or trace["quality"] != "exact"
                or trace["exact"] is not True):
            raise AuditError(f"{prefix} exactness contract drift")
        result.append({
            "algorithm": trace["algorithm"],
            "quality": trace["quality"],
            "exact": trace["exact"],
            "key_components": key,
            "matrix_cells": matrix,
            "side_a_size": size_a,
            "side_b_size": size_b,
            "smaller_side": expected_smaller,
            "total_cost": total,
            "source_pairs": [list(pair) for pair in sorted(observed_pairs)],
        })
    return result


def _inspect_publication_pair(
        label: str, formulas_path: Path, values_path: Path,
        witness_item: dict[str, object], expected: dict[str, object],
        workbook_identities: dict[str, dict[str, object]],
        ) -> tuple[dict[str, object], set[str]]:
    """Authenticate both schema-v3 envelopes and their external payload.

    This reader is audit-owned: it imports no product sidecar parser.  It binds
    each twin to the current workbook bytes and mtime, recomputes the shared
    generation binding, independently inflates every zlib member, and compares
    the complete persisted duplicate-pairing trace with the independent oracle.
    """
    paths = {"formulas": formulas_path, "values": values_path}
    envelopes: dict[str, dict[str, object]] = {}
    sidecars: dict[str, dict[str, object]] = {}
    for flavor, workbook in paths.items():
        sidecar = Path(str(workbook) + ".outcome.json")
        sentinel = Path(str(sidecar) + ".tmp")
        if sentinel.exists():
            raise AuditError(
                f"product {label} publication sentinel remains: "
                f"{sentinel.name}")
        if not sidecar.is_file() or sidecar.is_symlink():
            raise AuditError(
                f"product {label} sidecar is absent or indirect: "
                f"{sidecar.name}")
        raw = sidecar.read_bytes()
        if len(raw) > 1_048_576:
            raise AuditError(f"product {label} sidecar exceeds audit limit")
        try:
            envelope = json.loads(raw.decode("utf-8"))
        except (UnicodeError, json.JSONDecodeError) as exc:
            raise AuditError(
                f"product {label} sidecar is not strict UTF-8 JSON: {exc}") from exc
        if not isinstance(envelope, dict):
            raise AuditError(f"product {label} sidecar is not an object")
        if (
            envelope.get("schema_version"),
            envelope.get("comparison_schema_version"),
            envelope.get("record_type"),
            envelope.get("completion"),
            envelope.get("skipped_inputs"),
            envelope.get("failed_inputs"),
        ) != (1, 3, "comparison", "complete", 0, 0):
            raise AuditError(
                f"product {label} {flavor} sidecar state drift")
        generation = envelope.get("artifact_generation")
        if not isinstance(generation, dict) or (
                generation.get("completion"),
                generation.get("publication_state"),
                generation.get("requested_mode")) != (
                    "complete", "committed", "both"):
            raise AuditError(
                f"product {label} {flavor} generation state drift")
        generation_id = generation.get("generation_id")
        if not isinstance(generation_id, str) or not re.fullmatch(
                r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-"
                r"[0-9a-f]{4}-[0-9a-f]{12}", generation_id):
            raise AuditError(f"product {label} generation ID drift")
        members = generation.get("members")
        if not isinstance(members, list) or len(members) != 2:
            raise AuditError(f"product {label} twin member census drift")
        by_flavor = {
            str(member.get("flavor")): member
            for member in members if isinstance(member, dict)}
        if set(by_flavor) != set(paths):
            raise AuditError(f"product {label} twin member flavors drift")
        inspected_members = {
            member_flavor: _publication_member_exact(
                by_flavor[member_flavor], member_path, member_flavor,
                f"product {label} {flavor} envelope",
                workbook_identities[member_flavor])
            for member_flavor, member_path in paths.items()}
        if envelope.get("self_member") != by_flavor[flavor]:
            raise AuditError(f"product {label} {flavor} self-member drift")
        content_digests = generation.get("content_digests")
        if content_digests != {
                member_flavor: facts["sha256"]
                for member_flavor, facts in inspected_members.items()}:
            raise AuditError(f"product {label} generation digest map drift")
        if not isinstance(generation.get("producer_versions"), dict):
            raise AuditError(f"product {label} producer version map drift")
        manifest = envelope.get("comparison_payload")
        if not isinstance(manifest, dict):
            raise AuditError(f"product {label} payload manifest is absent")
        envelopes[flavor] = envelope
        sidecars[flavor] = {
            "path": str(sidecar.resolve()),
            "bytes": len(raw),
            "sha256": _sha_bytes(raw),
            "self_member_exact": True,
        }

    formula_envelope = envelopes["formulas"]
    values_envelope = envelopes["values"]
    generation = formula_envelope["artifact_generation"]
    manifest = formula_envelope["comparison_payload"]
    if generation != values_envelope.get("artifact_generation"):
        raise AuditError(f"product {label} twins declare different generations")
    if manifest != values_envelope.get("comparison_payload"):
        raise AuditError(f"product {label} twins declare different payloads")
    for key in ("completion", "skipped_inputs", "failed_inputs"):
        if formula_envelope.get(key) != values_envelope.get(key):
            raise AuditError(
                f"product {label} twins disagree on {key}")

    if (manifest.get("schema_version"), manifest.get("encoding")) != (
            1, "canonical-json-zlib-chunks-v1"):
        raise AuditError(f"product {label} payload schema drift")
    chunks = manifest.get("chunks")
    if not isinstance(chunks, list) or not chunks:
        raise AuditError(f"product {label} payload chunk manifest drift")
    referenced: set[str] = set()
    decoded_parts: list[bytes] = []
    chunk_evidence: list[dict[str, object]] = []
    parent = formulas_path.parent.resolve()
    for ordinal, descriptor in enumerate(chunks):
        if not isinstance(descriptor, dict):
            raise AuditError(f"product {label} payload descriptor drift")
        relative = descriptor.get("relative_path")
        if (not isinstance(relative, str)
                or not PAYLOAD_BASENAME_RE.fullmatch(relative)
                or Path(relative).name != relative
                or relative in referenced):
            raise AuditError(f"product {label} payload path drift: {relative!r}")
        referenced.add(relative)
        chunk = formulas_path.parent / relative
        if (not chunk.is_file() or chunk.is_symlink()
                or chunk.resolve().parent != parent):
            raise AuditError(
                f"product {label} payload chunk is absent or indirect")
        raw = chunk.read_bytes()
        if (len(raw) != descriptor.get("size")
                or _sha_bytes(raw) != descriptor.get("sha256")
                or len(raw) > 67_108_864):
            raise AuditError(
                f"product {label} payload chunk identity drift")
        inflater = zlib.decompressobj()
        try:
            decoded = inflater.decompress(raw) + inflater.flush()
        except zlib.error as exc:
            raise AuditError(
                f"product {label} payload chunk cannot inflate: {exc}") from exc
        if (not inflater.eof or inflater.unused_data or inflater.unconsumed_tail
                or len(decoded) != descriptor.get("decoded_size")):
            raise AuditError(
                f"product {label} payload chunk framing drift")
        decoded_parts.append(decoded)
        chunk_evidence.append({
            "ordinal": ordinal, "relative_path": relative,
            "bytes": len(raw), "sha256": _sha_bytes(raw),
            "decoded_bytes": len(decoded),
        })
    decoded = b"".join(decoded_parts)
    if (len(decoded) != manifest.get("decoded_size")
            or _sha_bytes(decoded) != manifest.get("decoded_sha256")
            or len(decoded) > 67_108_864):
        raise AuditError(f"product {label} decoded payload identity drift")
    try:
        persisted = json.loads(decoded.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as exc:
        raise AuditError(
            f"product {label} decoded payload is invalid JSON: {exc}") from exc
    if not isinstance(persisted, dict) or _canonical(persisted) != decoded:
        raise AuditError(f"product {label} payload is not canonical JSON")

    binding = _sha_bytes(_canonical({
        "decoded_sha256": manifest.get("decoded_sha256"),
        "completion": formula_envelope.get("completion"),
        "skipped_inputs": formula_envelope.get("skipped_inputs"),
        "failed_inputs": formula_envelope.get("failed_inputs"),
        "artifact_generation": generation,
    }))
    if manifest.get("binding_sha256") != binding:
        raise AuditError(f"product {label} payload generation binding drift")

    expected_counts = expected["counts"]
    persisted_counts = _normalized_payload_counts(
        persisted.get("counts"), f"product {label}")
    expected_verdict = (
        "match" if not expected_counts["differing_cells"]
        and not expected_counts["side_a_only_rows"]
        and not expected_counts["side_b_only_rows"] else "diff")
    if (
        persisted.get("status"), persisted.get("completion"),
        persisted.get("verdict"), persisted.get("pairing_quality"),
    ) != ("ok", "complete", expected_verdict, "exact"):
        raise AuditError(f"product {label} persisted outcome state drift")
    if persisted_counts != expected_counts:
        raise AuditError(f"product {label} persisted counts drift")
    expected_trace = expected["_pairing_trace_semantics"]
    trace = persisted.get("pairing_trace")
    persisted_trace_semantics = _persisted_trace_semantics(
        trace, f"product {label}")
    if (persisted_trace_semantics != expected_trace
            or persisted.get("duplicate_group_count") != len(expected_trace)
            or _sha_bytes(_canonical(persisted_trace_semantics))
            != expected["pairing_trace_sha256"]):
        raise AuditError(f"product {label} persisted pairing trace drift")
    for key in (
            "capped_group_diagnostics", "coverage_diagnostics",
            "failures", "warnings"):
        if persisted.get(key) != []:
            raise AuditError(
                f"product {label} persisted {key} is not empty")
    source_identities = persisted.get("source_identities")
    if not isinstance(source_identities, list):
        raise AuditError(f"product {label} source identity payload drift")

    returned = witness_item.get("result")
    if not isinstance(returned, dict):
        raise AuditError(f"product {label} witness return is absent")
    returned_generation = returned.get("artifact_generation")
    if not isinstance(returned_generation, dict):
        raise AuditError(f"product {label} returned generation is absent")
    returned_members = returned_generation.get("members")
    if not isinstance(returned_members, list) or len(returned_members) != 2:
        raise AuditError(f"product {label} returned member census drift")
    returned_by_flavor = {
        str(member.get("flavor")): member for member in returned_members
        if isinstance(member, dict)}
    generation_members = {
        str(member.get("flavor")): member
        for member in generation["members"] if isinstance(member, dict)}
    for flavor, path in paths.items():
        returned_member = returned_by_flavor.get(flavor)
        persisted_member = generation_members.get(flavor)
        if not isinstance(returned_member, dict) or not isinstance(
                persisted_member, dict):
            raise AuditError(f"product {label} member flavor binding drift")
        if returned_member != {
                "flavor": flavor,
                "commit_role": persisted_member["commit_role"],
                "path": str(path.resolve()),
                "bytes": persisted_member["size"],
                "sha256": persisted_member["sha256"]}:
            raise AuditError(
                f"product {label} returned/persisted member disagreement")

    return ({
        "sidecars": sidecars,
        "schema_version": 1,
        "comparison_schema_version": 3,
        "generation_id": generation["generation_id"],
        "generation_binding_sha256": binding,
        "generation_members": inspected_members,
        "generation_members_exact": True,
        "twin_generation_and_payload_manifests_identical": True,
        "payload_manifest": {
            "schema_version": manifest["schema_version"],
            "encoding": manifest["encoding"],
            "binding_sha256": manifest["binding_sha256"],
            "decoded_bytes": len(decoded),
            "decoded_sha256": _sha_bytes(decoded),
            "chunks": chunk_evidence,
        },
        "persisted_outcome": {
            "status": persisted["status"],
            "completion": persisted["completion"],
            "verdict": persisted["verdict"],
            "counts": persisted_counts,
            "pairing_quality": persisted["pairing_quality"],
            "duplicate_group_count": persisted["duplicate_group_count"],
            "pairing_trace_sha256": _sha_bytes(
                _canonical(persisted_trace_semantics)),
            "persisted_pairing_trace_wire_sha256": _sha_bytes(
                _canonical(trace)),
            "independent_pairing_trace_wire_sha256": expected[
                "independent_pairing_trace_wire_sha256"],
            "production_trace_internal_contract_exact": True,
            "source_identity_count": len(source_identities),
            "source_identities_sha256": _sha_bytes(
                _canonical(source_identities)),
            "diagnostic_lists_empty": True,
        },
        "persisted_counts_and_pairing_match_independent_oracle": True,
        "returned_generation_matches_persisted_generation": True,
    }, referenced)


def _expected_product_sheets(
        side_a: str, side_b: str, *, notes: bool, report_view: bool
        ) -> list[str]:
    names = [
        "Summary", "Spot Check", "Comparison", "Routes",
        f"Only in {side_a}", f"Only in {side_b}", side_a, side_b,
    ]
    if notes:
        names.append("Notes")
    if report_view:
        names.append("Report View")
    names.extend(["__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B"])
    return names


def _inspect_only_sheet(
        worksheet, *, present: str, missing: str,
        expected: Counter[tuple[object, ...]]) -> dict[str, object]:
    physical = worksheet.iter_rows(values_only=True)
    header = tuple(next(physical, ()))
    expected_header = (
        "Route", "Post Mile", "#", f"{present} Row",
        f"Missing from {missing}",
        *SHARED_FIELDS)
    if header != expected_header:
        raise AuditError(f"{worksheet.title}: one-sided header drift: {header!r}")
    observed: Counter[tuple[object, ...]] = Counter()
    for row in physical:
        if not row or all(value is None for value in row):
            continue
        if len(row) != len(expected_header):
            raise AuditError(f"{worksheet.title}: one-sided row width drift")
        observed[(
            _product_text(row[0]), _product_text(row[1]),
            *(_product_text(value) for value in row[5:]),
        )] += 1
    if observed != expected:
        raise AuditError(
            f"{worksheet.title}: one-sided inventory drift: "
            f"{_counter_difference_examples(observed, expected)!r}")
    return {
        "rows": sum(observed.values()),
        "ledger_sha256": _counter_digest(observed),
        "expected_ledger_sha256": _counter_digest(expected),
        "inventory_exact": True,
    }


def _inspect_product_workbook(
        label: str, formulas_path: Path, values_path: Path,
        expected: dict[str, object], left: Sequence[DetailRow],
        right: Sequence[DetailRow], *, notes: bool,
        report_view: str | None, tsn_rows: Sequence[DetailRow]
        ) -> dict[str, object]:
    side_a = str(expected["side_a"])
    side_b = str(expected["side_b"])
    expected_sheets = _expected_product_sheets(
        side_a, side_b, notes=notes, report_view=report_view is not None)
    formula_census = _formula_tag_census(formulas_path)
    value_formula_census = _formula_tag_census(values_path)
    expected_value_census = {
        "Summary": 12 if report_view is not None else 11,
        "Spot Check": 293,
        "Comparison": (
            2 * expected["counts"]["paired_rows"]
            + expected["counts"]["side_a_only_rows"]
            + expected["counts"]["side_b_only_rows"]),
        "Routes": 0,
        f"Only in {side_a}": expected["counts"]["side_a_only_rows"],
        f"Only in {side_b}": expected["counts"]["side_b_only_rows"],
        side_a: 2 * len(left) + 1,
        side_b: 2 * len(right) + 1,
        "__CMP_E2_SNAPSHOT_A": 0,
        "__CMP_E2_SNAPSHOT_B": 0,
    }
    if notes:
        expected_value_census["Notes"] = 0
    if report_view is not None:
        expected_value_census["Report View"] = 0
    route_count = len({row.route for row in (*left, *right)})
    expected_formula_census = {
        "Summary": 58 if report_view is not None else 57,
        "Spot Check": 293,
        "Comparison": 39 * (
            expected["counts"]["paired_rows"]
            + expected["counts"]["side_a_only_rows"]
            + expected["counts"]["side_b_only_rows"]),
        "Routes": 7 * route_count,
        f"Only in {side_a}": 36 * expected["counts"]["side_a_only_rows"],
        f"Only in {side_b}": 36 * expected["counts"]["side_b_only_rows"],
        side_a: 7 * len(left) + 1,
        side_b: 7 * len(right) + 1,
        "__CMP_E2_SNAPSHOT_A": 0,
        "__CMP_E2_SNAPSHOT_B": 0,
    }
    if notes:
        expected_formula_census["Notes"] = 0
    if report_view is not None:
        expected_formula_census["Report View"] = 0
    fixed_value_counts_exact = value_formula_census == expected_value_census
    fixed_formula_counts_exact = formula_census == expected_formula_census
    formula_contract = _formula_flavor_contract(
        formula_census, value_formula_census, expected_sheets,
        expected_formula_census, expected_value_census, side_a, side_b)
    if not formula_contract:
        raise AuditError(
            f"product {label} formula/value flavor contract drift: "
            f"formulas={formula_census!r} values={value_formula_census!r} "
            f"expected_formula_fixed={expected_formula_census!r} "
            f"expected_value_fixed={expected_value_census!r}")

    formula_book = load_workbook(
        formulas_path, read_only=True, data_only=False)
    try:
        if formula_book.sheetnames != expected_sheets:
            raise AuditError(
                f"product {label} formulas sheet universe drift: "
                f"{formula_book.sheetnames!r}")
    finally:
        formula_book.close()

    values = load_workbook(values_path, read_only=True, data_only=True)
    try:
        if values.sheetnames != expected_sheets:
            raise AuditError(
                f"product {label} values sheet universe drift: "
                f"{values.sheetnames!r}")
        if (values["__CMP_E2_SNAPSHOT_A"].sheet_state != "veryHidden"
                or values["__CMP_E2_SNAPSHOT_B"].sheet_state != "veryHidden"):
            raise AuditError(f"product {label} snapshots are not veryHidden")

        comparison = values["Comparison"]
        physical = comparison.iter_rows(values_only=True)
        header = tuple(next(physical, ()))
        state_header = (
            f"__CMP_E1_STATE_V1_C001_P0000_P{len(SHARED_FIELDS) - 1:04d}")
        expected_header = (
            "Route", "Post Mile", "#", f"{side_a} Row", f"{side_b} Row",
            "Status", "Diffs", *SHARED_FIELDS, state_header)
        if header != expected_header:
            raise AuditError(
                f"product {label} Comparison header drift: {header!r}")
        observed_pairs: Counter[tuple[object, ...]] = Counter()
        status_counts: Counter[str] = Counter()
        per_field: Counter[str] = Counter()
        differing_rows = differing_cells = union_rows = 0
        for row in physical:
            if not row or all(value is None for value in row):
                continue
            if len(row) != len(expected_header):
                raise AuditError(f"product {label} Comparison row width drift")
            union_rows += 1
            status = str(row[5])
            status_counts[status] += 1
            if status != "Both":
                continue
            state = str(row[-1])
            if len(state) != len(SHARED_FIELDS) or set(state) - {"E", "D"}:
                raise AuditError(f"product {label} paired state-vector drift")
            diffs = state.count("D")
            if row[6] != diffs:
                raise AuditError(f"product {label} Diffs/state disagreement")
            displays = tuple(_product_text(value) for value in row[7:-1])
            observed_pairs[(
                _product_text(row[0]), _product_text(row[1]), diffs,
                *displays, state,
            )] += 1
            differing_rows += int(diffs > 0)
            differing_cells += diffs
            for field, code in zip(SHARED_FIELDS, state):
                per_field[field] += int(code == "D")
        expected_pairs = expected["_paired"]
        if observed_pairs != expected_pairs:
            raise AuditError(
                f"product {label} paired cell ledger drift: "
                f"{_counter_difference_examples(observed_pairs, expected_pairs)!r}")
        expected_statuses = {
            "Both": expected["counts"]["paired_rows"],
            f"{side_a} only": expected["counts"]["side_a_only_rows"],
            f"{side_b} only": expected["counts"]["side_b_only_rows"],
        }
        expected_statuses = {
            key: value for key, value in expected_statuses.items()
            if value or key == "Both"}
        if dict(status_counts) != expected_statuses:
            raise AuditError(
                f"product {label} status census drift: {status_counts!r}")
        nonzero_per_field = {
            field: per_field[field] for field in SHARED_FIELDS
            if per_field[field]}
        if (union_rows != sum(expected_statuses.values())
                or differing_rows != expected["counts"]["differing_rows"]
                or differing_cells != expected["counts"]["differing_cells"]
                or nonzero_per_field != expected["counts"]["per_field_counts"]):
            raise AuditError(f"product {label} independently read counts drift")

        only_a = _inspect_only_sheet(
            values[f"Only in {side_a}"], present=side_a, missing=side_b,
            expected=expected["_side_a_only"])
        only_b = _inspect_only_sheet(
            values[f"Only in {side_b}"], present=side_b, missing=side_a,
            expected=expected["_side_b_only"])
        snapshot_a = _inspect_projection_sheet(
            values["__CMP_E2_SNAPSHOT_A"], left,
            side=side_a, snapshot=True)
        snapshot_b = _inspect_projection_sheet(
            values["__CMP_E2_SNAPSHOT_B"], right,
            side=side_b, snapshot=True)
        visible_a = _inspect_projection_sheet(
            values[side_a], left, side=side_a, snapshot=False)
        visible_b = _inspect_projection_sheet(
            values[side_b], right, side=side_b, snapshot=False)
        notes_contract = None
        if notes:
            note_text = "\n".join(
                str(value) for row in values["Notes"].iter_rows(values_only=True)
                for value in row if value is not None)
            notes_contract = {
                "declares_route_plus_canonical_postmile": (
                    "Rows are keyed on Route + the canonical Post Mile."
                    in note_text),
                "declares_county_identity": "County" in note_text,
                "declares_district_identity": "District" in note_text,
                "declares_source_dates_omitted": (
                    "REFERENCE_DATE" in note_text and "EXTRACT_DATE" in note_text),
            }
            if not notes_contract["declares_route_plus_canonical_postmile"]:
                raise AuditError(f"product {label} weak-key Notes drift")
        report_view_result = None
        if report_view is not None:
            report_view_result = _inspect_report_view(
                values["Report View"], expected_records=union_rows,
                side_a=side_a, tsn_rows=tsn_rows,
                raw_source_claims=report_view == "raw")
    finally:
        values.close()
    return {
        "formula_tag_census": formula_census,
        "values_formula_tag_census": value_formula_census,
        "fixed_formula_contract_exact": (
            fixed_value_counts_exact and fixed_formula_counts_exact),
        "formula_value_flavors_structurally_exact": True,
        "sheet_universe": expected_sheets,
        "counts": {
            "union_rows": union_rows,
            "paired_rows": status_counts["Both"],
            "side_a_only_rows": expected["counts"]["side_a_only_rows"],
            "side_b_only_rows": expected["counts"]["side_b_only_rows"],
            "differing_rows": differing_rows,
            "differing_cells": differing_cells,
            "per_field_counts": nonzero_per_field,
        },
        "paired_cell_ledger_sha256": _counter_digest(observed_pairs),
        "expected_paired_cell_ledger_sha256": expected[
            "paired_ledger_sha256"],
        "comparison_header": list(header),
        "comparison_identity_columns": ["Route", "Post Mile"],
        "district_column_present": "District" in header,
        "county_column_present": "County" in header,
        "only_in": {side_a: only_a, side_b: only_b},
        "snapshots": {side_a: snapshot_a, side_b: snapshot_b},
        "visible_source_sheets": {side_a: visible_a, side_b: visible_b},
        "notes_contract": notes_contract,
        "report_view": report_view_result,
    }


def _loaded_product_manifest_current(
        manifest: dict[str, object]) -> dict[str, object]:
    entries = manifest.get("entries")
    if not isinstance(entries, list) or manifest.get("file_count") != len(entries):
        raise AuditError("product module manifest shape drift")
    canonical = json.dumps(
        entries, sort_keys=True, separators=(",", ":")).encode("utf-8")
    if hashlib.sha256(canonical).hexdigest() != manifest.get(
            "canonical_json_sha256"):
        raise AuditError("product module manifest canonical digest drift")
    detail = []
    for entry in entries:
        if not isinstance(entry, dict):
            raise AuditError("product module manifest entry drift")
        path = REPO_ROOT / "scripts" / str(entry.get("relative_path", ""))
        observed = ({"bytes": path.stat().st_size,
                     "sha256": _sha_file(path)} if path.is_file() else None)
        expected = {"bytes": entry.get("bytes"),
                    "sha256": entry.get("sha256")}
        detail.append({
            "relative_path": entry.get("relative_path"),
            "expected": expected, "observed": observed,
            "current": observed == expected,
        })
    if not detail or not all(item["current"] for item in detail):
        raise AuditError("loaded product code changed after witness execution")
    return {
        "file_count": len(detail),
        "canonical_json_sha256": manifest.get("canonical_json_sha256"),
        "all_current": True,
        "entries": detail,
    }


def _product_loaded_raw_rows(
        rows: Sequence[DetailRow]) -> list[DetailRow]:
    """Independent twin of the current raw-TSN product projection.

    The source oracle retains exact decimal Length.  Production still passes
    that claim through binary64 before printing three decimals (CMP-AUD-138),
    so this projection is intentionally kept separate and compared back to the
    source-authoritative rows.
    """
    result = []
    length_index = SHARED_FIELDS.index("Length")
    raw_length_index = RAW_HEADERS.index("LENGTH")
    for row in rows:
        values = list(row.values)
        literal = _text(row.raw_values[raw_length_index])
        if literal:
            try:
                values[length_index] = f"{float(literal):07.3f}"
            except ValueError:
                values[length_index] = literal
        result.append(replace(
            row, source="product-loaded TSN raw",
            source_ref=f"product raw projection:{row.source_index + 2}",
            values=tuple(values)))
    return result


def _product_loaded_normalized_rows(
        rows: Sequence[DetailRow]) -> list[DetailRow]:
    """Independent twin of the current normalized-TSN product projection.

    The adapter applies ``pm_suffix`` to the separate PS cell rather than to a
    Post Mile token and therefore emits blank PS for every normalized row
    (CMP-AUD-042).  This is evidence about production, never source truth.
    """
    result = []
    ps_index = SHARED_FIELDS.index("PS")
    for row in rows:
        values = list(row.values)
        values[ps_index] = ""
        result.append(replace(
            row, source="product-loaded TSN normalized",
            source_ref=f"product normalized projection:{row.source_index + 2}",
            values=tuple(values)))
    return result


def _inspect_product(
        product_root: Path, *, excel_truth: Sequence[DetailRow],
        pdf_truth: Sequence[DetailRow], raw_rows: Sequence[DetailRow],
        normalized_rows: Sequence[DetailRow]) -> dict[str, object]:
    stale = sorted(
        [*product_root.glob("*.tmp-*.xlsx"),
         *product_root.glob("product-witness-*.partial.json"),
         *product_root.glob("product-witness-failure.json"),
         *product_root.glob("*.outcome.json.tmp")],
        key=lambda path: path.name)
    if stale:
        raise AuditError(
            "product witness root contains rejected temporary/failure state: "
            f"{[path.name for path in stale]!r}")
    result_path = product_root / "product-witness-result.json"
    payload = _read_json(result_path)
    consolidations = payload.get("consolidations")
    comparisons = payload.get("comparisons")
    if not isinstance(consolidations, dict) or set(consolidations) != {
            "excel", "pdf"}:
        raise AuditError("product witness consolidation manifest drift")
    expected_labels = {
        "excel_vs_tsn_raw", "excel_vs_tsn_normalized",
        "pdf_vs_tsn_raw", "pdf_vs_tsn_normalized", "pdf_vs_excel"}
    if not isinstance(comparisons, dict) or set(comparisons) != expected_labels:
        raise AuditError("product witness comparison-leg universe drift")
    consolidation_paths = {
        "excel": product_root / "highway_detail_excel_consolidated.xlsx",
        "pdf": product_root / "highway_detail_pdf_consolidated.xlsx",
    }
    truth = {"excel": excel_truth, "pdf": pdf_truth}
    consolidation_inspection: dict[str, object] = {}
    actual: dict[str, list[DetailRow]] = {}
    for flavor in ("excel", "pdf"):
        item = consolidations[flavor]
        if not isinstance(item, dict) or (
                item.get("status"), item.get("completion"),
                item.get("skipped_inputs"), item.get("failed_inputs")) != (
                    "ok", "complete", 0, 0):
            raise AuditError(f"product {flavor} consolidation outcome drift")
        _validate_file_identity(
            item.get("output", {}), consolidation_paths[flavor],
            f"product {flavor} consolidation")
        inspection, actual_rows = _inspect_consolidated_against_source(
            consolidation_paths[flavor], truth[flavor],
            f"TSMIS {flavor.title()}")
        actual[flavor] = actual_rows
        consolidation_inspection[flavor] = {
            **inspection,
            "file_identity": _file_identity(consolidation_paths[flavor]),
        }

    product_raw = _product_loaded_raw_rows(raw_rows)
    product_normalized = _product_loaded_normalized_rows(normalized_rows)
    input_projection = {
        "raw_tsn_source_vs_product_loader": _product_expected(
            "raw TSN source vs product loader", "Source", "Product loader",
            raw_rows, product_raw),
        "normalized_tsn_source_vs_product_loader": _product_expected(
            "normalized TSN source vs product loader", "Source",
            "Product loader", normalized_rows, product_normalized),
    }
    leg_contract = {
        "excel_vs_tsn_raw": (
            "TSMIS", "TSN", actual["excel"], product_raw, True, "raw"),
        "excel_vs_tsn_normalized": (
            "TSMIS", "TSN", actual["excel"], product_normalized,
            True, "normalized"),
        "pdf_vs_tsn_raw": (
            "TSMIS (PDF)", "TSN", actual["pdf"], product_raw, True, None),
        "pdf_vs_tsn_normalized": (
            "TSMIS (PDF)", "TSN", actual["pdf"], product_normalized,
            True, None),
        "pdf_vs_excel": (
            "TSMIS (PDF)", "TSMIS (Excel)", actual["pdf"],
            actual["excel"], False, None),
    }
    expected: dict[str, dict[str, object]] = {}
    inspected: dict[str, object] = {}
    referenced_chunks: set[str] = set()
    expected_sidecars: set[str] = set()
    for label, (side_a, side_b, left, right, notes, report_view) in (
            leg_contract.items()):
        expected[label] = _product_expected(
            label, side_a, side_b, left, right)
        item = comparisons[label]
        if not isinstance(item, dict):
            raise AuditError(f"product {label} witness payload drift")
        outcome = _helper_outcome_exact(label, item, expected[label])
        outputs = item.get("outputs")
        if not isinstance(outputs, dict) or set(outputs) != {
                "formulas", "values"}:
            raise AuditError(f"product {label} output manifest drift")
        formulas = product_root / f"{label}.xlsx"
        values = product_root / f"{label} (values).xlsx"
        formula_identity = _validate_file_identity(
            outputs["formulas"], formulas, f"{label} formulas")
        values_identity = _validate_file_identity(
            outputs["values"], values, f"{label} values")
        publication, leg_chunks = _inspect_publication_pair(
            label, formulas, values, item, expected[label], {
                "formulas": formula_identity,
                "values": values_identity,
            })
        if referenced_chunks.intersection(leg_chunks):
            raise AuditError(
                f"product {label} unexpectedly reuses another leg's payload chunk")
        referenced_chunks.update(leg_chunks)
        expected_sidecars.update({
            formulas.name + ".outcome.json",
            values.name + ".outcome.json",
        })
        inspected[label] = {
            "independent_weak_product_expectation":
                _public_product_expected(expected[label]),
            "returned_outcome": outcome,
            "outputs": {
                "formulas": formula_identity,
                "values": values_identity,
            },
            "independent_publication_inspection": publication,
            "independent_workbook_inspection": _inspect_product_workbook(
                label, formulas, values, expected[label], left, right,
                notes=notes, report_view=report_view,
                tsn_rows=(product_raw if report_view == "raw"
                          else product_normalized)),
        }
    present_chunks = {
        entry.name for entry in product_root.iterdir()
        if entry.is_file() and PAYLOAD_BASENAME_RE.fullmatch(entry.name)}
    if present_chunks != referenced_chunks:
        raise AuditError(
            "product payload artifact set is not exact: "
            f"present={sorted(present_chunks)!r}, "
            f"referenced={sorted(referenced_chunks)!r}")
    present_sidecars = {
        entry.name for entry in product_root.glob("*.xlsx.outcome.json")
        if entry.is_file()}
    if present_sidecars != expected_sidecars:
        raise AuditError(
            "product comparison sidecar set is not exact: "
            f"present={sorted(present_sidecars)!r}, "
            f"expected={sorted(expected_sidecars)!r}")
    lock_path = product_root / ".tsmis-comparison-publication.lock"
    if (not lock_path.is_file() or lock_path.is_symlink()
            or lock_path.stat().st_size != 0):
        raise AuditError("product publication lock identity drift")
    loaded = payload.get("loaded_product_code")
    if not isinstance(loaded, dict):
        raise AuditError("product witness omitted loaded-code manifest")
    return {
        "witness_result": _file_identity(result_path),
        "consolidations": consolidation_inspection,
        "tsn_input_projection": {
            label: _public_product_expected(item)
            for label, item in input_projection.items()},
        "comparisons": inspected,
        "publication_artifact_set": {
            "payload_chunks": sorted(referenced_chunks),
            "payload_chunk_count": len(referenced_chunks),
            "comparison_sidecars": sorted(expected_sidecars),
            "comparison_sidecar_count": len(expected_sidecars),
            "artifact_set_exact": True,
            "publication_lock": _file_identity(lock_path),
        },
        "loaded_product_code": loaded,
        "loaded_product_code_current": _loaded_product_manifest_current(loaded),
        "temporary_or_failure_artifacts": [],
    }


def _run_self_gate() -> dict[str, object]:
    completed = subprocess.run(
        [sys.executable, str(SELF_GATE_PATH)], cwd=REPO_ROOT,
        text=True, capture_output=True, timeout=120, check=False)
    if completed.returncode != 0:
        raise AuditError(
            "Highway Detail Stage-8 mutation gate failed: "
            f"{completed.stderr or completed.stdout}")
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    try:
        payload = json.loads(lines[-1])
    except (IndexError, json.JSONDecodeError) as exc:
        raise AuditError("Stage-8 mutation gate emitted no JSON") from exc
    if payload != {"status": "pass", "assertions": 79}:
        raise AuditError(f"Stage-8 mutation gate result drift: {payload!r}")
    return {
        "status": "executed_pass", "assertions": 79,
        "stdout": completed.stdout.strip(),
        "gate": _file_identity(SELF_GATE_PATH),
    }


def _file_identity(path: Path) -> dict[str, object]:
    return {
        "path": str(path.resolve()),
        "bytes": path.stat().st_size,
        "sha256": _sha_file(path),
    }


def _audit_code_identities() -> dict[str, dict[str, object]]:
    return {
        label: _file_identity(path)
        for label, path in AUDIT_CODE_PATHS.items()}


def _runtime_dependencies() -> dict[str, object]:
    return {
        "python": {
            "executable": str(Path(sys.executable).resolve()),
            "version": sys.version,
        },
        "packages": {
            name: importlib.metadata.version(name)
            for name in ("openpyxl", "pdfplumber")
        },
    }


def _bind_file(label: str, path: Path) -> dict[str, object]:
    observed = _file_identity(path)
    expected = FILE_BINDINGS[label]
    if {key: observed[key] for key in ("bytes", "sha256")} != expected:
        raise AuditError(f"{label} identity drift: {observed!r}")
    return {"binding": dict(expected), "observed": observed}


def _tree_manifest(root: Path, suffix: str) -> dict[str, object]:
    entries = []
    for path in sorted(root.glob(f"*{suffix}"), key=lambda item: item.name):
        entries.append((path.name, path.stat().st_size, _sha_file(path)))
    wire = "".join(
        f"{name}\t{size}\t{digest}\n" for name, size, digest in entries
    ).encode("utf-8")
    return {
        "files": len(entries),
        "bytes": sum(size for _name, size, _digest in entries),
        "manifest_sha256": _sha_bytes(wire),
        "members": [{"name": name, "bytes": size, "sha256": digest}
                    for name, size, digest in entries],
    }


def _bind_historical_owner_tree(
        label: str, root: Path, suffix: str, *, exact_universe: bool
        ) -> dict[str, object]:
    expected = HISTORICAL_OWNER_TREE_BINDINGS[label]
    names = tuple(
        f"highway_detail_route_{member}{suffix}"
        for member in ("005", "005S"))
    entries = []
    for name in names:
        path = root / name
        if not path.is_file() or path.is_symlink():
            raise AuditError(
                f"historical 7.7 owner source missing/non-file: {path}")
        entries.append((name, path.stat().st_size, _sha_file(path)))
    wire = "".join(
        f"{name}\t{size}\t{digest}\n" for name, size, digest in entries
    ).encode("utf-8")
    observed = {
        "files": len(entries),
        "bytes": sum(size for _name, size, _digest in entries),
        "manifest_sha256": _sha_bytes(wire),
        "members": [{"name": name, "bytes": size, "sha256": digest}
                    for name, size, digest in entries],
    }
    if any(observed[field] != expected[field]
           for field in ("files", "bytes", "manifest_sha256")):
        raise AuditError(
            f"historical 7.7 owner {label} tree drift: {observed!r}")
    present = {
        path.name for path in root.glob(f"*{suffix}") if path.is_file()}
    if exact_universe and present != set(names):
        raise AuditError(
            "private historical owner capture member universe drift: "
            f"{sorted(present)!r}")
    return {
        "root": str(root.resolve()),
        "binding": {**expected, "suffix": suffix},
        "observed": observed,
        "selected_members": list(names),
        "selected_universe_exact": present == set(names),
    }


def _read_json(path: Path) -> dict[str, object]:
    try:
        result = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise AuditError(f"cannot read dependency {path}: {exc}") from exc
    if not isinstance(result, dict):
        raise AuditError(f"dependency is not a JSON object: {path}")
    return result


def _accepted_dependencies(args: argparse.Namespace) -> dict[str, object]:
    identities = {
        "tsn_raw": _bind_file("tsn_raw", args.tsn_raw),
        "tsn_normalized": _bind_file("tsn_normalized", args.tsn_normalized),
        "tsn_normalized_sidecar": _bind_file(
            "tsn_normalized_sidecar", args.tsn_normalized_sidecar),
        "stage6_result": _bind_file("stage6_result", args.stage6_result),
        "stage6_acceptance": _bind_file(
            "stage6_acceptance", args.stage6_acceptance),
        "tsn_cross_format": _bind_file(
            "tsn_cross_format", args.tsn_cross_format),
    }
    pdf_tree = _tree_manifest(args.tsn_pdf_root, ".pdf")
    if {key: pdf_tree[key] for key in TSN_PDF_TREE_BINDING} != TSN_PDF_TREE_BINDING:
        raise AuditError("TSN PDF evidence tree identity drift")

    sidecar = _read_json(args.tsn_normalized_sidecar)
    sidecar_checks = {
        "complete": (
            sidecar.get("completion"), sidecar.get("skipped_inputs"),
            sidecar.get("failed_inputs")) == ("complete", 0, 0),
        "normalization_version_2": sidecar.get("tsn_normalization_version") == 2,
        "raw_member_bound": sidecar.get("tsn_raw_manifest", {}).get("members") == [{
            "relative_path": "TSAR - HIGHWAY DETAIL_TSN.xlsx",
            "byte_length": FILE_BINDINGS["tsn_raw"]["bytes"],
            "sha256": FILE_BINDINGS["tsn_raw"]["sha256"],
        }],
        "normalized_member_bound": (
            sidecar.get("tsn_normalized_workbook_identity", {}).get("byte_length"),
            sidecar.get("tsn_normalized_workbook_identity", {}).get("sha256"),
        ) == (
            FILE_BINDINGS["tsn_normalized"]["bytes"],
            FILE_BINDINGS["tsn_normalized"]["sha256"],
        ),
    }
    if not all(sidecar_checks.values()):
        raise AuditError(f"normalized sidecar contract drift: {sidecar_checks!r}")

    stage6 = _read_json(args.stage6_result)
    acceptance = _read_json(args.stage6_acceptance)
    projection = stage6.get("projection_comparison", {})
    stage6_checks = {
        "accepted_result_bound": (
            acceptance.get("accepted") is True
            and acceptance.get("result", {}).get("bytes")
            == FILE_BINDINGS["stage6_result"]["bytes"]
            and acceptance.get("result", {}).get("sha256")
            == FILE_BINDINGS["stage6_result"]["sha256"]),
        "postwrite_revalidated": (
            acceptance.get("post_result_write_revalidation") is True),
        "23_invariants_true": (
            len(stage6.get("audit_invariants", {})) == 23
            and all(stage6.get("audit_invariants", {}).values())),
        "family_complete": stage6.get("stage6_family_audit_complete") is True,
        "one_classified_projection_defect": (
            stage6.get("projection_exact") is False
            and stage6.get("unexplained_projection_residue_count") == 0
            and projection.get("typed_cell_mismatch_count") == 1
            and projection.get("typed_cell_mismatches_by_field") == {"Length": 1}),
        "known_full_conservation_red": (
            stage6.get("normalized_full_conservation") is False
            and len(stage6.get("findings", {}).get("blocking", [])) == 3),
    }
    if not all(stage6_checks.values()):
        raise AuditError(f"Stage-6 dependency contract drift: {stage6_checks!r}")

    cross = _read_json(args.tsn_cross_format)
    reconciliation = cross.get("reconciliation", {})
    cross_checks = {
        "status_pass": cross.get("status") == "pass",
        "fatal_reasons_empty": cross.get("fatal_reasons") == [],
        "60081_of_60083_paired": (
            reconciliation.get("xlsx_records"),
            reconciliation.get("pdf_records"),
            reconciliation.get("paired_records"),
            reconciliation.get("xlsx_only_records"),
            reconciliation.get("pdf_only_records"),
        ) == (60_083, 60_081, 60_081, 2, 0),
        "all_residue_classified": (
            reconciliation.get("source_date_delta_cells") == 441
            and cross.get("delta_allowlist", {}).get("matches") is True
            and cross.get("delta_allowlist", {}).get("expected_item_count") == 443),
        "report_view_map_present": bool(cross.get("row_to_report_view_mapping")),
    }
    if not all(cross_checks.values()):
        raise AuditError(f"TSN cross-format dependency drift: {cross_checks!r}")
    return {
        "identities": identities,
        "tsn_pdf_tree": {"binding": TSN_PDF_TREE_BINDING, "observed": pdf_tree},
        "normalized_sidecar": {"checks": sidecar_checks},
        "stage6_raw_to_normalized": {
            "checks": stage6_checks,
            "projection_comparison": projection,
            "classified_projection_residue": stage6.get(
                "classified_projection_residue"),
            "identity_and_collision_census": stage6.get(
                "identity_and_collision_census"),
            "blocking_findings": stage6.get("findings", {}).get("blocking"),
        },
        "tsn_xlsx_to_pdf": {
            "checks": cross_checks,
            "source_snapshot_relation": cross.get("source_snapshot_relation"),
            "reconciliation": reconciliation,
            "delta_allowlist": cross.get("delta_allowlist"),
            "field_to_pdf_mapping": cross.get("field_to_pdf_mapping"),
            "row_to_report_view_mapping": cross.get("row_to_report_view_mapping"),
        },
    }


def run(args: argparse.Namespace) -> dict[str, object]:
    mutation_gate = _run_self_gate()
    indexed_equivalence = _indexed_oracle_equivalence_gate()
    provenance = {
        "tsmis_excel": tsmis_source._bind_tree(
            "tsmis_excel", args.tsmis_xlsx_root),
        "tsmis_pdf": tsmis_source._bind_tree(
            "tsmis_pdf", args.tsmis_pdf_root),
        "historical_owner_excel": _bind_historical_owner_tree(
            "xlsx", args.historical_owner_xlsx_root, ".xlsx",
            exact_universe=True),
        "historical_owner_pdf": _bind_historical_owner_tree(
            "pdf", args.historical_owner_pdf_root, ".pdf",
            exact_universe=True),
    }
    live_origins = {
        "tsmis_excel": tsmis_source._bind_tree(
            "tsmis_excel", args.origin_tsmis_xlsx_root),
        "tsmis_pdf": tsmis_source._bind_tree(
            "tsmis_pdf", args.origin_tsmis_pdf_root),
        "tsn_raw": _bind_file("tsn_raw", args.origin_tsn_raw),
        "historical_owner_excel": _bind_historical_owner_tree(
            "xlsx", args.origin_historical_owner_xlsx_root, ".xlsx",
            exact_universe=False),
        "historical_owner_pdf": _bind_historical_owner_tree(
            "pdf", args.origin_historical_owner_pdf_root, ".pdf",
            exact_universe=False),
    }
    origin_tsn_pdf = _tree_manifest(args.origin_tsn_pdf_root, ".pdf")
    if {key: origin_tsn_pdf[key] for key in TSN_PDF_TREE_BINDING
            } != TSN_PDF_TREE_BINDING:
        raise AuditError("live-origin TSN PDF tree identity drift")
    live_origins["tsn_pdf"] = {
        "binding": TSN_PDF_TREE_BINDING, "observed": origin_tsn_pdf}
    dependencies = _accepted_dependencies(args)

    excel_source_rows, excel_source_summary = tsmis_source._parse_excel(
        args.tsmis_xlsx_root, None)
    pdf_source_rows, pdf_source_summary = tsmis_source._parse_pdf(
        args.tsmis_pdf_root, None, args.workers)
    if pdf_source_summary["unclassified_groups"]:
        raise AuditError("TSMIS PDF source parse has unclassified groups")
    format_alignment = tsmis_source._source_format_alignment(
        excel_source_rows, pdf_source_rows)
    tsmis_source_contract = {
        "excel_rows": len(excel_source_rows)
        == EXPECTED_TSMIS_SOURCE["excel_rows"],
        "excel_ordered_typed_rows": excel_source_summary[
            "ordered_typed_rows_sha256"]
        == EXPECTED_TSMIS_SOURCE["excel_ordered_typed_rows_sha256"],
        "pdf_rows": len(pdf_source_rows)
        == EXPECTED_TSMIS_SOURCE["pdf_rows"],
        "pdf_ordered_typed_rows": pdf_source_summary[
            "ordered_typed_rows_sha256"]
        == EXPECTED_TSMIS_SOURCE["pdf_ordered_typed_rows_sha256"],
        "pdf_reconciliation": pdf_source_summary["reconciliation"]
        == EXPECTED_TSMIS_SOURCE["pdf_reconciliation"],
        "format_totals": format_alignment["totals"]
        == EXPECTED_TSMIS_SOURCE["format_totals"],
        "format_pair_map": format_alignment["pair_map_sha256"]
        == EXPECTED_TSMIS_SOURCE["format_pair_map_sha256"],
    }
    if not all(tsmis_source_contract.values()):
        raise AuditError(
            f"frozen current TSMIS source contract drift: "
            f"{tsmis_source_contract!r}")
    format_pairs, pair_map = _format_pairs(
        excel_source_rows, pdf_source_rows, format_alignment)
    attested_excel, excel_attestation = _attest_excel_county(
        excel_source_rows, pdf_source_rows, format_pairs)

    historical_excel_rows, historical_excel_summary = (
        tsmis_source._parse_excel(args.historical_owner_xlsx_root, None))
    historical_pdf_rows, historical_pdf_summary = tsmis_source._parse_pdf(
        args.historical_owner_pdf_root, None, args.workers)
    if historical_pdf_summary["unclassified_groups"]:
        raise AuditError(
            "historical 7.7 TSMIS PDF owner source has unclassified groups")
    owner_members = {"005", "005S"}
    current_owner_excel_rows = [
        row for row in excel_source_rows if row.member_route in owner_members]
    historical_alignment = tsmis_source._source_format_alignment(
        current_owner_excel_rows, historical_pdf_rows)
    historical_pairs, historical_pair_map = _format_pairs(
        current_owner_excel_rows, historical_pdf_rows, historical_alignment)
    historical_exact_pairs = [
        item for item in historical_pairs
        if (item[0].member_route == "005"
            and item[2] == "all_34_render_equal")]
    historical_attested_excel, historical_excel_attestation = (
        _attest_excel_county(
            current_owner_excel_rows, historical_pdf_rows,
            historical_exact_pairs))
    historical_self_alignment = tsmis_source._source_format_alignment(
        historical_excel_rows, historical_pdf_rows)
    historical_source_contract = {
        "excel_rows": len(historical_excel_rows)
        == EXPECTED_HISTORICAL_OWNER_SOURCE["excel_rows"],
        "excel_ordered_typed_rows": historical_excel_summary[
            "ordered_typed_rows_sha256"]
        == EXPECTED_HISTORICAL_OWNER_SOURCE[
            "excel_ordered_typed_rows_sha256"],
        "pdf_rows": len(historical_pdf_rows)
        == EXPECTED_HISTORICAL_OWNER_SOURCE["pdf_rows"],
        "pdf_ordered_typed_rows": historical_pdf_summary[
            "ordered_typed_rows_sha256"]
        == EXPECTED_HISTORICAL_OWNER_SOURCE[
            "pdf_ordered_typed_rows_sha256"],
        "pdf_reconciliation": historical_pdf_summary["reconciliation"]
        == EXPECTED_HISTORICAL_OWNER_SOURCE["pdf_reconciliation"],
        "current_excel_alignment_totals": historical_alignment["totals"]
        == EXPECTED_HISTORICAL_OWNER_SOURCE[
            "current_excel_to_historical_pdf_totals"],
        "current_excel_alignment_pair_map": historical_alignment[
            "pair_map_sha256"]
        == EXPECTED_HISTORICAL_OWNER_SOURCE[
            "current_excel_to_historical_pdf_pair_map_sha256"],
        "historical_self_pair_map": historical_self_alignment[
            "pair_map_sha256"]
        == EXPECTED_HISTORICAL_OWNER_SOURCE[
            "historical_excel_to_pdf_pair_map_sha256"],
        "same_build_route_005_owner_rows": len(historical_attested_excel)
        == EXPECTED_HISTORICAL_OWNER_SOURCE[
            "same_build_route_005_owner_rows"],
        "route_005_excel_bytes_identical_across_captures": (
            _sha_file(args.tsmis_xlsx_root /
                      "highway_detail_route_005.xlsx")
            == _sha_file(args.historical_owner_xlsx_root /
                         "highway_detail_route_005.xlsx")
            == EXPECTED_HISTORICAL_OWNER_SOURCE[
                "route_005_current_and_historical_excel_sha256"]),
    }
    if not all(historical_source_contract.values()):
        raise AuditError(
            "frozen historical owner-source contract drift: "
            f"{historical_source_contract!r}")

    excel_rows = [
        _from_tsmis_source(row, district="", county="")
        for row in excel_source_rows]
    pdf_rows = [_from_tsmis_source(row) for row in pdf_source_rows]
    raw_rows, raw_summary = _parse_tsn_raw(args.tsn_raw)
    normalized_rows, normalized_summary = _parse_tsn_normalized(
        args.tsn_normalized)
    constrained_excel, excel_owner_constraints = (
        _analyze_excel_owner_constraints(
            excel_source_rows, pdf_source_rows, raw_rows, attested_excel,
            historical_attested_excel))
    constrained_excel_summary = _source_summary(
        "TSMIS Excel rows with exact or unanimous companion-PDF owner",
        constrained_excel, {
            "owner_constraint_analysis": excel_owner_constraints})
    owner_constraint_comparisons = {
        "snapshot_attested_excel_vs_tsn_raw": _comparison(
            "snapshot-attested TSMIS Excel vs TSN raw",
            constrained_excel, raw_rows),
        "snapshot_attested_excel_vs_tsn_normalized": _comparison(
            "snapshot-attested TSMIS Excel vs TSN normalized",
            constrained_excel, normalized_rows),
    }
    if any(item["pairing_quality"] != "exact"
           or item["capped_diagnostics"]
           for item in owner_constraint_comparisons.values()):
        raise AuditError("Excel owner-constraint comparison did not complete exactly")
    owner_constraint_sha256 = _sha_bytes(_canonical(excel_owner_constraints))
    owner_comparison_contracts = {
        label: _sha_bytes(_canonical({
            "counts": item["counts"],
            "ordered_pair_ledger_sha256": item[
                "ordered_pair_ledger_sha256"],
        })) for label, item in owner_constraint_comparisons.items()}
    if (owner_constraint_sha256 != EXPECTED_OWNER_CONSTRAINT_SHA256
            or owner_comparison_contracts
            != EXPECTED_OWNER_COMPARISON_CONTRACT_SHA256):
        raise AuditError(
            "frozen snapshot-owner comparison contract drift: "
            f"{owner_constraint_sha256} / {owner_comparison_contracts!r}")
    attested_excel_summary = _source_summary(
        "TSMIS Excel rows with uniquely printed companion owner",
        attested_excel, {"attestation": excel_attestation})
    pdf_summary = _source_summary(
        "TSMIS PDF", pdf_rows, {
            "source_reconstruction": pdf_source_summary,
        })

    comparisons = {
        "raw_vs_normalized": _comparison(
            "TSN raw vs normalized", raw_rows, normalized_rows),
        "pdf_vs_tsn_raw": _comparison(
            "TSMIS PDF vs TSN raw", pdf_rows, raw_rows),
        "pdf_vs_tsn_normalized": _comparison(
            "TSMIS PDF vs TSN normalized", pdf_rows, normalized_rows),
        "excel_attested_vs_tsn_raw": _comparison(
            "uniquely owner-attested TSMIS Excel vs TSN raw",
            attested_excel, raw_rows),
        "excel_attested_vs_tsn_normalized": _comparison(
            "uniquely owner-attested TSMIS Excel vs TSN normalized",
            attested_excel, normalized_rows),
    }
    for label, comparison in comparisons.items():
        if (comparison["counts"] != EXPECTED_SOURCE_COUNTS[label]
                or comparison["ordered_pair_ledger_sha256"]
                != EXPECTED_SOURCE_LEDGERS[label]):
            raise AuditError(
                f"{label}: frozen source comparison drift: "
                f"{comparison['counts']!r} / "
                f"{comparison['ordered_pair_ledger_sha256']}")
    production = _inspect_product(
        args.product_root, excel_truth=excel_rows, pdf_truth=pdf_rows,
        raw_rows=raw_rows, normalized_rows=normalized_rows)
    production_contract_sha256 = _sha_bytes(_canonical(production))
    if production_contract_sha256 != EXPECTED_PRODUCT_OBJECT_SHA256:
        raise AuditError(
            "frozen complete production red-result contract drift: "
            f"{production_contract_sha256}")
    product_books = {
        label: item["independent_workbook_inspection"]
        for label, item in production["comparisons"].items()}
    product_publications = {
        label: item["independent_publication_inspection"]
        for label, item in production["comparisons"].items()}

    raw_normalized_counts = comparisons["raw_vs_normalized"]["counts"]
    invariants = {
        "exact_tsmis_trees_rebound": (
            provenance["tsmis_excel"]["observed"]["files"] == 252
            and provenance["tsmis_pdf"]["observed"]["files"] == 252),
        "private_captures_match_live_origins": (
            provenance["tsmis_excel"]["observed"]
            == live_origins["tsmis_excel"]["observed"]
            and provenance["tsmis_pdf"]["observed"]
            == live_origins["tsmis_pdf"]["observed"]
            and provenance["historical_owner_excel"]["observed"]
            == live_origins["historical_owner_excel"]["observed"]
            and provenance["historical_owner_pdf"]["observed"]
            == live_origins["historical_owner_pdf"]["observed"]
            and live_origins["tsn_raw"]["observed"]["bytes"]
            == FILE_BINDINGS["tsn_raw"]["bytes"]
            and live_origins["tsn_raw"]["observed"]["sha256"]
            == FILE_BINDINGS["tsn_raw"]["sha256"]
            and {key: live_origins["tsn_pdf"]["observed"][key]
                 for key in TSN_PDF_TREE_BINDING}
            == TSN_PDF_TREE_BINDING),
        "accepted_tsn_dependencies_rebound": all(
            all(section["checks"].values()) for section in (
                dependencies["normalized_sidecar"],
                dependencies["stage6_raw_to_normalized"],
                dependencies["tsn_xlsx_to_pdf"],
            )),
        "tsmis_pdf_zero_unclassified": (
            pdf_source_summary["unclassified_groups"] == 0),
        "tsmis_source_rows_exact": (
            len(excel_source_rows), len(pdf_source_rows)) == (51_273, 51_216),
        "frozen_tsmis_source_contract_exact": all(
            tsmis_source_contract.values()),
        "historical_owner_sources_separately_bound_and_exact": (
            all(historical_source_contract.values())
            and historical_pair_map["matches_public_source_oracle"]
            and historical_excel_attestation["county_inference_used"]
            is False),
        "format_pair_map_reproduced": pair_map[
            "matches_public_source_oracle"],
        "excel_county_never_inferred": (
            excel_attestation["county_inference_used"] is False
            and len(attested_excel)
            + excel_attestation["total_county_unknown_excel_rows"] == 51_273),
        "excel_owner_constraint_residue_fully_classified": (
            sum(excel_owner_constraints["classifications"].values()) == 51_273
            and excel_owner_constraints["owner_attested_rows"]
            == len(constrained_excel) == 51_273
            and excel_owner_constraints["unresolved_owner_rows"] == 0
            and excel_owner_constraints[
                "same_build_historical_owner_rows_promoted"] == 3_125
            and excel_owner_constraints[
                "current_exact_rows_superseded_by_same_build_snapshot"]
            == 2_608
            and excel_owner_constraints["new_companion_key_constrained_rows"]
            == 3
            and excel_owner_constraints["cross_edition_owner_conflicts"] == 8
            and excel_owner_constraints["composite_description_attestation"][
                "uniquely_owner_attested_rows"] == 2
            and excel_owner_constraints["composite_description_attestation"][
                "ambiguous_eligible_rows"] == 0
            and excel_owner_constraints["composite_description_attestation"][
                "row_equivalence_claimed"] is False
            and excel_owner_constraints["tsn_only_owner_promotions"] == 0
            and all(item["pairing_quality"] == "exact"
                    for item in owner_constraint_comparisons.values())),
        "snapshot_owner_and_full_excel_comparisons_frozen_exact": (
            owner_constraint_sha256 == EXPECTED_OWNER_CONSTRAINT_SHA256
            and owner_comparison_contracts
            == EXPECTED_OWNER_COMPARISON_CONTRACT_SHA256),
        "tsn_raw_and_normalized_rows_exact": (
            len(raw_rows), len(normalized_rows)) == (60_083, 60_083),
        "raw_normalized_one_known_length_defect_only": (
            raw_normalized_counts["paired_rows"] == 60_083
            and raw_normalized_counts["side_a_only_rows"] == 0
            and raw_normalized_counts["side_b_only_rows"] == 0
            and raw_normalized_counts["differing_rows"] == 1
            and raw_normalized_counts["differing_cells"] == 1
            and raw_normalized_counts["per_field_counts"] == {"Length": 1}),
        "all_pairing_exact_not_capped": all(
            item["pairing_quality"] == "exact"
            and not item["capped_diagnostics"]
            for item in comparisons.values()),
        "every_comparison_reconciles_side_counts": all(
            item["counts"]["paired_rows"] + item["counts"]["side_a_only_rows"]
            == ({
                "raw_vs_normalized": len(raw_rows),
                "pdf_vs_tsn_raw": len(pdf_rows),
                "pdf_vs_tsn_normalized": len(pdf_rows),
                "excel_attested_vs_tsn_raw": len(attested_excel),
                "excel_attested_vs_tsn_normalized": len(attested_excel),
            }[label])
            and item["counts"]["paired_rows"]
            + item["counts"]["side_b_only_rows"] == 60_083
            for label, item in comparisons.items()),
        "all_source_counts_and_ledgers_match_frozen_oracle": all(
            item["counts"] == EXPECTED_SOURCE_COUNTS[label]
            and item["ordered_pair_ledger_sha256"]
            == EXPECTED_SOURCE_LEDGERS[label]
            for label, item in comparisons.items()),
        "indexed_oracle_semantic_equivalence": indexed_equivalence["exact"],
        "permanent_mutation_gate_executed_pass": (
            mutation_gate["status"] == "executed_pass"),
        "production_excel_consolidation_source_exact": production[
            "consolidations"]["excel"]["source_projection_exact"],
        "complete_production_red_result_matches_frozen_contract": (
            production_contract_sha256 == EXPECTED_PRODUCT_OBJECT_SHA256),
        "production_pdf_consolidation_known_red_reproduced": not production[
            "consolidations"]["pdf"]["source_projection_exact"],
        "production_raw_loader_one_length_defect_reproduced": (
            production["tsn_input_projection"][
                "raw_tsn_source_vs_product_loader"]["counts"] == {
                    "known": True, "paired_rows": 60_083,
                    "side_a_only_rows": 0, "side_b_only_rows": 0,
                    "differing_rows": 1, "differing_cells": 1,
                    "per_field_counts": {"Length": 1},
                    "asserted_cells": 2_042_822, "context_cells": 0,
                }),
        "production_normalized_loader_ps_loss_reproduced": (
            production["tsn_input_projection"][
                "normalized_tsn_source_vs_product_loader"]["counts"] == {
                    "known": True, "paired_rows": 60_083,
                    "side_a_only_rows": 0, "side_b_only_rows": 0,
                    "differing_rows": 1_177, "differing_cells": 1_177,
                    "per_field_counts": {"PS": 1_177},
                    "asserted_cells": 2_042_822, "context_cells": 0,
                }),
        "production_all_five_formula_value_legs_self_consistent": (
            len(product_books) == 5
            and all(item["formula_value_flavors_structurally_exact"]
                    for item in product_books.values())),
        "production_all_five_publications_cryptographically_exact": (
            len(product_publications) == 5
            and all(item[
                "persisted_counts_and_pairing_match_independent_oracle"]
                    and item["returned_generation_matches_persisted_generation"]
                    and item[
                        "twin_generation_and_payload_manifests_identical"]
                    for item in product_publications.values())
            and production["publication_artifact_set"][
                "payload_chunk_count"] == 5
            and production["publication_artifact_set"][
                "comparison_sidecar_count"] == 10),
        "production_comparison_source_provenance_absence_reproduced": all(
            item["persisted_outcome"]["source_identity_count"] == 0
            for item in product_publications.values()),
        "production_all_five_paired_cell_ledgers_exact": all(
            item["paired_cell_ledger_sha256"]
            == item["expected_paired_cell_ledger_sha256"]
            for item in product_books.values()),
        "production_all_twenty_source_views_and_snapshots_exact": all(
            all(view["projection_exact"] for view in (
                *item["snapshots"].values(),
                *item["visible_source_sheets"].values()))
            for item in product_books.values()),
        "production_all_one_sided_inventories_exact": all(
            all(side["inventory_exact"] for side in item["only_in"].values())
            for item in product_books.values()),
        "production_raw_report_view_maps_all_tsn_source_claims": (
            product_books["excel_vs_tsn_raw"]["report_view"][
                "source_only_mapping_exact"] is True),
        "production_normalized_report_view_reproduces_source_claim_omission": (
            not any(product_books["excel_vs_tsn_normalized"]["report_view"][
                "source_only_nonblank_counts"].values())),
        "production_pdf_tsn_report_view_absence_reproduced": (
            product_books["pdf_vs_tsn_raw"]["report_view"] is None
            and product_books["pdf_vs_tsn_normalized"]["report_view"] is None),
        "production_weak_route_pm_identity_explicit": all(
            item["comparison_identity_columns"] == ["Route", "Post Mile"]
            and not item["district_column_present"]
            and not item["county_column_present"]
            for item in product_books.values()),
        "loaded_product_code_current": production[
            "loaded_product_code_current"]["all_current"],
    }
    if not all(invariants.values()):
        raise AuditError(f"source-comparison invariants failed: {invariants!r}")

    stage8_complete = all(invariants.values())
    production_tsmis_projection_exact = all(
        item["source_projection_exact"]
        for item in production["consolidations"].values())
    production_overlapping_cells_exact = all(
        item["paired_cell_ledger_sha256"]
        == item["expected_paired_cell_ledger_sha256"]
        for item in product_books.values())
    route_005_alignment = next(
        item for item in format_alignment["per_member"]
        if item["member"] == "005")
    route_395_pdf_source = next(
        item for item in pdf_source_summary["per_file"]
        if item["member"] == "395")
    pdf_projection = production["consolidations"]["pdf"]
    findings = {
        "oracle_blocking": [],
        "source_export_deltas": [{
            "finding": "CMP-AUD-192",
            "classification": "cross-edition source snapshot mismatch",
            "fact": (
                "The route-005 Excel member in the current folder is byte-"
                "identical to the 7.7 Excel member and matches the 7.7 PDF "
                "at all 3,125 rows. The later PDF is a different snapshot; "
                "eight unchanged rows carry different printed DCR owners."),
            "evidence": {
                "route_005_current_format_alignment": route_005_alignment,
                "same_build_alignment": historical_alignment["per_member"][0],
                "owner_conflicts": excel_owner_constraints[
                    "cross_edition_owner_conflicts_exact"],
                "owner_conflict_ledger_sha256": excel_owner_constraints[
                    "cross_edition_owner_conflict_ledger_sha256"],
            },
        }, {
            "finding": "CMP-AUD-191",
            "classification": "printed composite source mapping",
            "fact": (
                "The two final 005S Excel rows are not separate PDF records, "
                "but their unique exact Descriptions are visibly printed as "
                "components of one 07/LA composite PDF row. Only owner is "
                "attested; no PDF row or non-owner cells are synthesized."),
            "evidence": excel_owner_constraints[
                "composite_description_attestation"],
        }],
        "product_red": [{
            "finding": "CMP-AUD-042",
            "fact": (
                "The normalized Highway Detail product loader blanks PS on "
                "exactly 1,177 rows; the raw product loader does not."),
            "evidence": production["tsn_input_projection"][
                "normalized_tsn_source_vs_product_loader"],
        }, {
            "finding": "CMP-AUD-045",
            "fact": (
                "All five persisted product workbooks key only on Route and "
                "Post Mile and expose neither District nor County, while the "
                "source oracle uses snapshot-backed County plus complete PP, "
                "numeric Post Mile, and roadbed."),
            "evidence": {
                label: {
                    "identity_columns": item["comparison_identity_columns"],
                    "district_present": item["district_column_present"],
                    "county_present": item["county_column_present"],
                } for label, item in product_books.items()},
        }, {
            "finding": "CMP-AUD-054",
            "fact": (
                "The live product PDF consolidation is not the audit-owned "
                "current source projection: statewide it has 34 source-only "
                "and 24 product-only raw rows, including route-005 residue."),
            "evidence": {
                "difference_counts": pdf_projection[
                    "raw_multiset_difference_counts"],
                "source_only_ledger_sha256": pdf_projection[
                    "source_only_raw_ledger_sha256"],
                "product_only_ledger_sha256": pdf_projection[
                    "product_only_raw_ledger_sha256"],
                "examples": pdf_projection["raw_multiset_difference_examples"],
            },
        }, {
            "finding": "CMP-AUD-186",
            "fact": (
                "The audit source parser exactly retains route 395's proven "
                "multi-baseline line-two record; the production projection "
                "still has two source-only and two product-only route-395 rows."),
            "evidence": {
                "source_reconstruction": route_395_pdf_source,
                "production_route_395": {
                    "source_only": pdf_projection[
                        "raw_multiset_difference_counts"][
                            "source_only_by_route"].get("395", 0),
                    "product_only": pdf_projection[
                        "raw_multiset_difference_counts"][
                            "product_only_by_route"].get("395", 0),
                },
            },
        }, {
            "finding": "CMP-AUD-068",
            "fact": (
                "Both PDF-vs-TSN product legs omit Report View; the two "
                "Excel-vs-TSN legs include it."),
            "evidence": {
                label: item["sheet_universe"]
                for label, item in product_books.items()},
        }, {
            "finding": "CMP-AUD-076",
            "fact": (
                "Every cryptographically valid persisted comparison payload "
                "contains zero durable source identities."),
            "evidence": {
                label: item["persisted_outcome"]["source_identity_count"]
                for label, item in product_publications.items()},
        }, {
            "finding": "CMP-AUD-133",
            "fact": (
                "The raw Excel-vs-TSN Report View maps source claims, while "
                "the normalized Report View reproduces their complete omission."),
            "evidence": {
                "raw": product_books["excel_vs_tsn_raw"]["report_view"],
                "normalized": product_books[
                    "excel_vs_tsn_normalized"]["report_view"],
            },
        }, {
            "finding": "CMP-AUD-138",
            "fact": (
                "The raw product loader reproduces the one exact Length cell "
                "changed by binary64 conversion."),
            "evidence": production["tsn_input_projection"][
                "raw_tsn_source_vs_product_loader"],
        }, {
            "finding": "CMP-AUD-142",
            "fact": (
                "Product Notes explicitly omit REFERENCE_DATE and "
                "EXTRACT_DATE instead of preserving the printed snapshot "
                "claims needed to distinguish these source editions."),
            "evidence": {
                label: item["notes_contract"]
                for label, item in product_books.items()
                if item["notes_contract"] is not None},
        }],
        "audit_harness_remediated": [
            "CMP-AUD-188", "CMP-AUD-189", "CMP-AUD-190", "CMP-AUD-191"],
    }

    return {
        "schema_version": 1,
        "audit": "Stage 8 Highway Detail authoritative four-source oracle",
        "status": "complete" if stage8_complete else "incomplete",
        "current_stage": (
            "complete source reconstruction, snapshot-backed physical truth, "
            "independent production-workbook inspection, frozen red-result "
            "contract, and detached post-write publication decision"),
        "methodology": {
            "authority": (
                "Current 252-member TSMIS Excel/PDF trees, separately bound "
                "same-build 7.7 route-005/005S owner sources, authoritative "
                "raw TSN XLSX and twelve district PDFs, accepted normalized "
                "TSN, Stage-6 conservation, and XLSX/PDF mapping."),
            "snapshot_policy": (
                "Folder labels never blend editions. Route 005 uses its "
                "byte-identical 7.7 Excel/PDF pair for Excel ownership; the "
                "later PDF remains a separate source snapshot. 005S uses the "
                "current PDF and explicit composite component evidence."),
            "independence": (
                "Source truth imports no application parser, comparator, "
                "consolidator, schema, writer, or workbook result. Production "
                "ran in an isolated witness and every persisted artifact was "
                "independently re-read."),
        },
        "provenance": provenance,
        "live_origin_provenance": live_origins,
        "dependencies": dependencies,
        "indexed_oracle_equivalence_gate": indexed_equivalence,
        "dependency_gates": {
            "stage8_highway_detail_mutations": mutation_gate},
        "sources": {
            "tsmis_excel_raw_reconstruction": excel_source_summary,
            "tsmis_pdf": pdf_summary,
            "tsmis_excel_pdf_source_alignment": format_alignment,
            "tsmis_frozen_source_contract": tsmis_source_contract,
            "tsmis_excel_owner_attestation": attested_excel_summary,
            "tsmis_excel_owner_constraints": constrained_excel_summary,
            "historical_7_7_owner_corroboration": {
                "scope": (
                    "separately versioned route 005/005S TSMIS Excel/PDF "
                    "sources; route 005 is selected only because its Excel "
                    "member is byte-identical and all 3,125 historical PDF "
                    "rows are exact; 005S is corroboration only"),
                "historical_excel": historical_excel_summary,
                "historical_pdf": historical_pdf_summary,
                "current_excel_to_historical_pdf_alignment":
                    historical_alignment,
                "historical_excel_to_historical_pdf_alignment":
                    historical_self_alignment,
                "same_build_route_005_owner_attestation":
                    historical_excel_attestation,
                "pair_map_reproduction": historical_pair_map,
                "frozen_contract": historical_source_contract,
            },
            "tsn_raw": raw_summary,
            "tsn_normalized": normalized_summary,
        },
        "physical_identity": {
            "definition": [
                "route including route suffix", "printed/source County",
                "complete PP", "numeric Post Mile", "effective roadbed"],
            "equation_marker_role": (
                "separately asserted PS field; not a pairing component"),
            "district_role": (
                "separately asserted owner claim; not a pairing component"),
            "excel_limitation": (
                "no County exists in the 34-column workbook package; only a "
                "uniquely owned current signature, exact separately versioned "
                "TSMIS signature over the current row payload, or exact unique "
                "component of a printed current composite may attest it"),
        },
        "source_comparisons": comparisons,
        "source_owner_constraint_comparisons": owner_constraint_comparisons,
        "source_owner_constraint_contracts": {
            "owner_constraint_sha256": owner_constraint_sha256,
            "comparison_contract_sha256": owner_comparison_contracts,
        },
        "production": production,
        "production_object_sha256": production_contract_sha256,
        "findings": findings,
        "audit_invariants": invariants,
        "source_oracle_exact": True,
        "source_truth_exact": True,
        "production_tsmis_projection_exact":
            production_tsmis_projection_exact,
        "production_overlapping_comparison_cells_exact":
            production_overlapping_cells_exact,
        "production_value_projection_exact": False,
        "production_comparison_semantics_exact": False,
        "excel_vs_tsn_full_physical_truth_known": (
            excel_owner_constraints["unresolved_owner_rows"] == 0
            and len(constrained_excel) == 51_273),
        "stage8_base_oracle_complete": stage8_complete,
        "statewide_acceptance": stage8_complete,
        "comparison_end_to_end_perfect": False,
        "code_provenance": {
            "audit_code": _audit_code_identities(),
            "runtime_dependencies": _runtime_dependencies(),
            "application_modules_imported_into_truth_process": [],
        },
    }


def _atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_bytes(payload)
    temporary.replace(path)


def _unlink_if_present(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass


def _identity_current(expected: dict[str, object]) -> bool:
    path = Path(str(expected.get("path", "")))
    if not path.is_file() or path.is_symlink():
        return False
    actual = _file_identity(path)
    return all(actual.get(key) == expected.get(key)
               for key in ("path", "bytes", "sha256"))


def _publication_current(
        args: argparse.Namespace, result: dict[str, object]
        ) -> tuple[bool, dict[str, object]]:
    """Re-read every decisive input and emitted evidence after result commit."""
    try:
        source_current = {
            "tsmis_excel": tsmis_source._bind_tree(
                "tsmis_excel", args.tsmis_xlsx_root),
            "tsmis_pdf": tsmis_source._bind_tree(
                "tsmis_pdf", args.tsmis_pdf_root),
            "historical_owner_excel": _bind_historical_owner_tree(
                "xlsx", args.historical_owner_xlsx_root, ".xlsx",
                exact_universe=True),
            "historical_owner_pdf": _bind_historical_owner_tree(
                "pdf", args.historical_owner_pdf_root, ".pdf",
                exact_universe=True),
        }
        origin_current = {
            "tsmis_excel": tsmis_source._bind_tree(
                "tsmis_excel", args.origin_tsmis_xlsx_root),
            "tsmis_pdf": tsmis_source._bind_tree(
                "tsmis_pdf", args.origin_tsmis_pdf_root),
            "tsn_raw": _bind_file("tsn_raw", args.origin_tsn_raw),
            "historical_owner_excel": _bind_historical_owner_tree(
                "xlsx", args.origin_historical_owner_xlsx_root, ".xlsx",
                exact_universe=False),
            "historical_owner_pdf": _bind_historical_owner_tree(
                "pdf", args.origin_historical_owner_pdf_root, ".pdf",
                exact_universe=False),
        }
        origin_pdf = _tree_manifest(args.origin_tsn_pdf_root, ".pdf")
        if {key: origin_pdf[key] for key in TSN_PDF_TREE_BINDING
                } != TSN_PDF_TREE_BINDING:
            raise AuditError("post-write live TSN PDF tree drift")
        origin_current["tsn_pdf"] = {
            "binding": TSN_PDF_TREE_BINDING, "observed": origin_pdf}
        dependency_current = _accepted_dependencies(args)
        code_current = _audit_code_identities()
        runtime_current = _runtime_dependencies()

        source_checks = {
            label: source_current[label] == result["provenance"][label]
            for label in source_current}
        origin_checks = {
            label: origin_current[label]
            == result["live_origin_provenance"][label]
            for label in origin_current}
        dependency_checks = {
            "complete_dependency_contract": (
                dependency_current == result["dependencies"])}
        code_checks = {
            label: identity
            == result["code_provenance"]["audit_code"][label]
            for label, identity in code_current.items()}
        runtime_checks = {
            "runtime_dependencies": runtime_current
            == result["code_provenance"]["runtime_dependencies"]}

        production = result["production"]
        artifact_checks: dict[str, bool] = {
            "witness_result": _identity_current(
                production["witness_result"])}
        for flavor, consolidation in production["consolidations"].items():
            artifact_checks[f"consolidation:{flavor}"] = _identity_current(
                consolidation["file_identity"])
        for label, comparison in production["comparisons"].items():
            for flavor, identity in comparison["outputs"].items():
                artifact_checks[f"workbook:{label}:{flavor}"] = (
                    _identity_current(identity))
            publication = comparison["independent_publication_inspection"]
            for flavor, identity in publication["sidecars"].items():
                artifact_checks[f"sidecar:{label}:{flavor}"] = (
                    _identity_current(identity))
            for flavor, member in publication["generation_members"].items():
                path = Path(str(member["path"]))
                artifact_checks[f"generation-member:{label}:{flavor}"] = (
                    path.is_file() and not path.is_symlink()
                    and path.stat().st_mtime_ns == member["mtime_ns"]
                    and path.stat().st_size == member["size"]
                    and _sha_file(path) == member["sha256"])
            for chunk in publication["payload_manifest"]["chunks"]:
                path = args.product_root / str(chunk["relative_path"])
                expected = {
                    "path": str(path.resolve()), "bytes": chunk["bytes"],
                    "sha256": chunk["sha256"],
                }
                artifact_checks[
                    f"payload:{label}:{chunk['ordinal']}"] = (
                        _identity_current(expected))
        lock_expected = production["publication_artifact_set"][
            "publication_lock"]
        artifact_checks["publication_lock"] = _identity_current(lock_expected)

        expected_chunks = set(production["publication_artifact_set"][
            "payload_chunks"])
        present_chunks = {
            path.name for path in args.product_root.iterdir()
            if path.is_file() and PAYLOAD_BASENAME_RE.fullmatch(path.name)}
        expected_sidecars = set(production["publication_artifact_set"][
            "comparison_sidecars"])
        present_sidecars = {
            path.name for path in args.product_root.glob("*.xlsx.outcome.json")
            if path.is_file()}
        artifact_set_checks = {
            "payload_chunk_set_exact": present_chunks == expected_chunks,
            "comparison_sidecar_set_exact": (
                present_sidecars == expected_sidecars),
            "no_temporary_partial_or_failure_artifacts": not any((
                *args.product_root.glob("*.tmp-*.xlsx"),
                *args.product_root.glob("product-witness-*.partial.json"),
                *args.product_root.glob("product-witness-failure.json"),
                *args.product_root.glob("*.outcome.json.tmp"),
            )),
        }
        loaded = _loaded_product_manifest_current(
            production["loaded_product_code"])
        groups = (
            source_checks, origin_checks, dependency_checks, code_checks,
            runtime_checks, artifact_checks, artifact_set_checks)
        current = all(value for group in groups for value in group.values())
        current = current and loaded["all_current"]
        return current, {
            "source_checks": source_checks,
            "origin_checks": origin_checks,
            "dependency_checks": dependency_checks,
            "code_checks": code_checks,
            "runtime_checks": runtime_checks,
            "artifact_checks": artifact_checks,
            "artifact_set_checks": artifact_set_checks,
            "loaded_product_code_current": loaded["all_current"],
        }
    except Exception as exc:
        return False, {
            "revalidation_error": {
                "type": type(exc).__name__, "message": str(exc)}}


def _write_decision(
        path: Path, output: Path, result: dict[str, object], *,
        accepted: bool, reason: str, postwrite_current: bool,
        postwrite_detail: dict[str, object], open_findings_authorized: bool,
        ) -> dict[str, object]:
    identity = _file_identity(output)
    decision = {
        "schema_version": 1, "accepted": accepted, "reason": reason,
        "audit": result.get("audit"), "result": str(output.resolve()),
        "result_bytes": identity["bytes"],
        "result_sha256": identity["sha256"],
        "source_truth_exact": result.get("source_truth_exact", False),
        "production_tsmis_projection_exact": result.get(
            "production_tsmis_projection_exact", False),
        "production_overlapping_comparison_cells_exact": result.get(
            "production_overlapping_comparison_cells_exact", False),
        "production_value_projection_exact": result.get(
            "production_value_projection_exact", False),
        "production_comparison_semantics_exact": result.get(
            "production_comparison_semantics_exact", False),
        "stage8_base_oracle_complete": result.get(
            "stage8_base_oracle_complete", False),
        "comparison_end_to_end_perfect": result.get(
            "comparison_end_to_end_perfect", False),
        "open_product_findings_authorized": open_findings_authorized,
        "post_result_write_revalidation": postwrite_current,
        "post_result_write_identities": postwrite_detail,
    }
    _atomic_write(path, (json.dumps(
        decision, ensure_ascii=False, sort_keys=True, indent=2
    ) + "\n").encode("utf-8"))
    return decision


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--tsmis-xlsx-root", type=Path,
                        default=DEFAULT_TSMIS_XLSX_ROOT)
    parser.add_argument("--tsmis-pdf-root", type=Path,
                        default=DEFAULT_TSMIS_PDF_ROOT)
    parser.add_argument("--historical-owner-xlsx-root", type=Path,
                        default=DEFAULT_HISTORICAL_OWNER_XLSX_ROOT)
    parser.add_argument("--historical-owner-pdf-root", type=Path,
                        default=DEFAULT_HISTORICAL_OWNER_PDF_ROOT)
    parser.add_argument("--origin-tsmis-xlsx-root", type=Path,
                        default=DEFAULT_ORIGIN_TSMIS_XLSX_ROOT)
    parser.add_argument("--origin-tsmis-pdf-root", type=Path,
                        default=DEFAULT_ORIGIN_TSMIS_PDF_ROOT)
    parser.add_argument(
        "--origin-historical-owner-xlsx-root", type=Path,
        default=DEFAULT_ORIGIN_HISTORICAL_OWNER_XLSX_ROOT)
    parser.add_argument(
        "--origin-historical-owner-pdf-root", type=Path,
        default=DEFAULT_ORIGIN_HISTORICAL_OWNER_PDF_ROOT)
    parser.add_argument("--tsn-raw", type=Path, default=DEFAULT_TSN_RAW)
    parser.add_argument("--origin-tsn-raw", type=Path,
                        default=DEFAULT_ORIGIN_TSN_RAW)
    parser.add_argument("--tsn-pdf-root", type=Path,
                        default=DEFAULT_TSN_PDF_ROOT)
    parser.add_argument("--origin-tsn-pdf-root", type=Path,
                        default=DEFAULT_ORIGIN_TSN_PDF_ROOT)
    parser.add_argument("--tsn-normalized", type=Path,
                        default=DEFAULT_TSN_NORMALIZED)
    parser.add_argument("--tsn-normalized-sidecar", type=Path,
                        default=DEFAULT_TSN_NORMALIZED_SIDECAR)
    parser.add_argument("--stage6-result", type=Path,
                        default=DEFAULT_STAGE6_RESULT)
    parser.add_argument("--stage6-acceptance", type=Path,
                        default=DEFAULT_STAGE6_ACCEPTANCE)
    parser.add_argument("--tsn-cross-format", type=Path,
                        default=DEFAULT_TSN_CROSS_FORMAT)
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--product-root", type=Path,
                        default=DEFAULT_PRODUCT_ROOT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--allow-open-findings", action="store_true",
        help=(
            "accept the complete source/product audit while the exact "
            "documented production findings remain open"))
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    acceptance_path = args.output.with_suffix(
        args.output.suffix + ".acceptance.json")
    rejection_path = args.output.with_suffix(
        args.output.suffix + ".rejection.json")
    _unlink_if_present(acceptance_path)
    _unlink_if_present(rejection_path)
    try:
        result = run(args)
    except Exception as exc:
        result = {
            "schema_version": 1,
            "audit": "Stage 8 Highway Detail authoritative four-source oracle",
            "status": "failed",
            "error": {"type": type(exc).__name__, "message": str(exc)},
            "source_truth_exact": False,
            "production_tsmis_projection_exact": False,
            "production_overlapping_comparison_cells_exact": False,
            "production_value_projection_exact": False,
            "production_comparison_semantics_exact": False,
            "stage8_base_oracle_complete": False,
            "comparison_end_to_end_perfect": False,
        }
        encoded = (json.dumps(
            result, ensure_ascii=False, sort_keys=True, indent=2
        ) + "\n").encode("utf-8")
        _atomic_write(args.output, encoded)
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False,
            reason="oracle_execution_failed", postwrite_current=False,
            postwrite_detail={}, open_findings_authorized=False)
        sys.stdout.write(json.dumps({
            "accepted": False, "reason": decision["reason"],
            "error": result["error"],
            "output": str(args.output.resolve()),
            "rejection": str(rejection_path.resolve()),
        }, ensure_ascii=False, separators=(",", ":")) + "\n")
        return 1

    encoded = (json.dumps(
        result, ensure_ascii=False, sort_keys=True, indent=2, default=str,
    ) + "\n").encode("utf-8")
    _atomic_write(args.output, encoded)
    postwrite_current, postwrite_detail = _publication_current(args, result)
    complete = result.get("stage8_base_oracle_complete") is True
    open_findings = bool(result.get("findings", {}).get("product_red"))
    accepted = bool(
        complete and postwrite_current
        and (not open_findings or args.allow_open_findings))
    if not complete:
        reason = "stage8_base_oracle_incomplete"
    elif not postwrite_current:
        reason = "post_result_write_revalidation_failed"
    elif open_findings and not args.allow_open_findings:
        reason = "open_product_findings_require_explicit_authorization"
    else:
        reason = "accepted_complete_audit_with_documented_open_product_findings"
    decision_path = acceptance_path if accepted else rejection_path
    decision = _write_decision(
        decision_path, args.output, result, accepted=accepted, reason=reason,
        postwrite_current=postwrite_current,
        postwrite_detail=postwrite_detail,
        open_findings_authorized=bool(args.allow_open_findings))
    sys.stdout.write(json.dumps({
        "status": result["status"], "accepted": decision["accepted"],
        "reason": decision["reason"],
        "output": str(args.output.resolve()),
        "bytes": len(encoded),
        "sha256": _sha_bytes(encoded),
        "decision": str(decision_path.resolve()),
        "invariants": (
            f"{sum(result['audit_invariants'].values())}/"
            f"{len(result['audit_invariants'])}"),
    }, separators=(",", ":")) + "\n")
    return 0 if accepted else 1


if __name__ == "__main__":
    raise SystemExit(main())
