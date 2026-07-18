"""Golden check for the cross-environment Intersection comparisons (v0.17.0):
compare_env.INTERSECTION_SUMMARY (AGGREGATE per route) + INTERSECTION_DETAIL (FLAT,
route+PM). Both promote their report to a full matrix row (Everything + by-day).

Locks:
  * wiring — both adapters registered in COMPARE_REPORTS (folders/env) and present in
    reports.matrix_rows(); the TSN-only extra-rows list is now EMPTY (every report has
    a cross-env adapter), so nothing is duplicated in the by-day matrix;
  * Summary AGGREGATE path — side_loader / sheet_name None / agg_header == IS_HEADER,
    route-keyed, a genuine per-route diff flagged (via a stub side-loader; the real
    block-walk parser is golden-tested in check_consolidate_intersection);
  * Detail FLAT path — sheet "Intersection Detail", key_col "Post Mile"; end-to-end on
    synthetic per-route XLSX that a non-key cell change is flagged while the key (PM)
    aligns the rows, with the environment names as the two sides.

CI-safe (no real files / network). Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_env_intersection.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
import reports
from events import Events
from openpyxl import Workbook, load_workbook

_fail = []
DIFF = " ≠ "


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def _summary(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        text = []
        for r in wb["Summary"].iter_rows(values_only=True):
            text += [str(c) for c in r if c is not None and str(c).strip()]
        return " ".join(text), wb.sheetnames
    finally:
        wb.close()


def _comparison_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        it = wb["Comparison"].iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        rows = [["" if c is None else str(c) for c in r] for r in it
                if r and any(c not in (None, "") for c in r)]
        return header, rows
    finally:
        wb.close()


def test_wiring():
    print("wiring (both Intersection cross-env adapters; extra rows empty):")
    s, d = compare_env.INTERSECTION_SUMMARY, compare_env.INTERSECTION_DETAIL
    check("Summary: aggregate path (side_loader, sheet_name None, agg_header IS_HEADER)",
          s.side_loader is compare_env._load_intersection_summary_side
          and s.sheet_name is None and s.agg_header is compare_env.IS_HEADER)
    check("Detail: flat path (sheet 'Intersection Detail', key_col 'Post Mile')",
          d.sheet_name == "Intersection Detail" and d.key_col == "Post Mile"
          and d.side_loader is None)
    env_adapters = {adapter for _l, adapter, kind, group in reports.COMPARE_REPORTS
                    if kind == "folders" and group == "env"}
    check("both registered as folders/env in COMPARE_REPORTS",
          s in env_adapters and d in env_adapters)
    mrows = {r[0] for r in reports.matrix_rows()}
    check("both are matrix rows", "intersection_summary" in mrows
          and "intersection_detail" in mrows)
    check("TSN-only extra-rows list is now empty (no by-day duplication)",
          reports.tsn_matrix_extra_rows() == [])


def test_summary_aggregate_compare():
    print("Summary AGGREGATE env compare (route-keyed; one genuine per-route diff):")
    H = compare_env.IS_HEADER
    ncat = len(H) - 2

    def mk(route, total, firstcat):
        return [route, total, firstcat] + [0] * (ncat - 1)

    def stub(folder, label, events):       # noqa: ARG001
        is_b = "ars" in str(folder).lower()
        return [mk("001", 10, 3), mk("002", 6 if is_b else 5, 1)], []

    adapter = compare_env.EnvCompare(
        "intersection_summary", "Intersection Summary", "intersection_summary",
        side_loader=stub, agg_header=H,
        base_schema=compare_env.INTERSECTION_SUMMARY.base_schema)
    root = Path(tempfile.mkdtemp(prefix="isenv_"))
    da, db = root / "2026-06-19 ssor-prod", root / "2026-06-19 ars-prod"
    da.mkdir(parents=True); db.mkdir(parents=True)
    out = root / "cmp.xlsx"
    res = adapter.compare_folders(str(da), str(db), str(out), events=Events(),
                                  confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    blob, sheets = _summary(out)
    check("side labels are the environments", "SSOR-PROD" in blob and "ARS-PROD" in blob)
    check("keyed on Route", "keyed on Route" in blob)
    check("exactly one differing cell (route 002 Total)",
          res.comparison_outcome is not None
          and res.comparison_outcome.counts.differing_cells == 1)
    check("env side sheets present",
          "SSOR-PROD" in sheets and "ARS-PROD" in sheets and "Comparison" in sheets)


def _write_id_route(folder, route, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Intersection Detail"
    ws.append(["P", "Post Mile", "S", "Location", "Ctrl Type", "Description"])
    for r in rows:
        ws.append(r)
    wb.save(folder / f"intersection_detail_route_{route}.xlsx")
    wb.close()


def test_detail_flat_compare():
    print("Detail FLAT env compare (route+PM key; a non-key cell diff flagged):")
    root = Path(tempfile.mkdtemp(prefix="idenv_"))
    a = root / "2026-06-19 ssor-prod" / "intersection_detail"
    b = root / "2026-06-19 ars-prod" / "intersection_detail"
    a.mkdir(parents=True); b.mkdir(parents=True)
    rows_a = [["R", "0.204", None, "12 ORA 001", "S", "JCT 5"],
              ["R", "1.000", None, "12 ORA 001", "S", "PT A"]]
    rows_b = [["R", "0.204", None, "12 ORA 001", "S", "JCT 5"],
              ["R", "1.000", None, "12 ORA 001", "F", "PT A"]]   # Ctrl Type S->F at PM 1.000
    _write_id_route(a, "001", rows_a)
    _write_id_route(b, "001", rows_b)
    out = root / "cmp.xlsx"
    res = compare_env.INTERSECTION_DETAIL.compare_folders(
        str(a.parent), str(b.parent), str(out), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    blob, sheets = _summary(out)
    check("keyed on Route + Post Mile", "keyed on Route + Post Mile" in blob)
    check("env side sheets present", "SSOR-PROD" in sheets and "ARS-PROD" in sheets)
    header, rows = _comparison_rows(out)
    check("Post Mile is the key column", "Post Mile" in header)
    pm = header.index("Post Mile")
    by = {r[pm]: r for r in rows}
    ctrl = header.index("Ctrl Type")
    # CMP-AUD-045: the Comparison key column shows the canonical identity
    # display (route / county / PP+Decimal-PM).
    check("both postmiles matched (county-aware keyed, no spurious one-sided)",
          "001 / ORA / R0.204" in by and "001 / ORA / R1" in by)
    check("PM 1.000 Ctrl Type S vs F is a genuine diff",
          DIFF in by["001 / ORA / R1"][ctrl])
    check("PM 0.204 unchanged row has no diff",
          DIFF not in " ".join(by["001 / ORA / R0.204"]))

    # CMP-AUD-076 (folder-kind): the committed comparison persists a durable
    # provenance record — full canonical FOLDER selections, roles = the derived
    # side labels, and the exact discovered member census per side.
    import compare_tsn_common as ctc
    prov = ctc.read_comparison_provenance(out)
    check("a provenance sidecar exists beside the env comparison", prov is not None)
    ins = prov.get("inputs") or []
    # Resolved-to-resolved (the CI runner's 8.3 short temp names expand).
    check("both sides recorded as folder-kind with full selections",
          [i.get("kind") for i in ins] == ["folder", "folder"]
          and str(a.parent.resolve()) in ins[0]["selection"]
          and str(b.parent.resolve()) in ins[1]["selection"])
    check("roles are the derived side labels",
          ins[0].get("role") == "SSOR-PROD" and ins[1].get("role") == "ARS-PROD")
    check("the exact discovered member census is recorded per side",
          all(i["member_count"] == 1
              and i["members"][0]["name"] == "intersection_detail_route_001.xlsx"
              and i["members"][0]["size"] > 0 for i in ins))
    check("the record binds the committed generation",
          prov.get("generation_id") == res.artifact_generation.generation_id
          and prov.get("recipe", {}).get("report") == "Intersection Detail")
    from openpyxl import load_workbook
    wb = load_workbook(out, read_only=True, data_only=True)
    try:
        flatp = " ".join(str(c) for row in wb["Provenance"].iter_rows(values_only=True)
                         for c in row if c is not None)
    finally:
        wb.close()
    check("a Provenance sheet shows folder selections + member counts",
          str(a.parent.resolve()) in flatp and str(b.parent.resolve()) in flatp
          and "1 discovered source file(s)" in flatp)


def _id_full_row(**over):
    """A full 35-column per-route Intersection Detail row (current site layout),
    defaulted so the pairing key resolves, with named overrides by column label."""
    import compare_intersection_detail_tsn as _idt
    body = list(_idt._TSMIS_HEADER[1:])          # 35 site columns
    row = [""] * len(body)
    base = {"PP": "R", "Post Mile": "1.000", "PS": "", "Location": "12 ORA 001",
            "Description": "MAIN ST"}
    for k, v in {**base, **over}.items():
        row[body.index(k)] = v
    return row


def _write_id_route_hdr(folder, route, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Intersection Detail"
    ws.append(list(header))
    for r in rows:
        ws.append(r)
    wb.save(folder / f"intersection_detail_route_{route}.xlsx")
    wb.close()


def test_detail_edition_boundary():
    """CMP-AUD (2026-07-17): the cross-env Intersection Detail comparison bridges
    the July-2026 LABEL-ONLY edition change — a current-labelled export compares
    against a legacy-labelled one by POSITION (values proven identical in place),
    instead of refusing with 'different column layouts'."""
    print("Detail edition boundary (current vs legacy labels align by position):")
    import compare_intersection_detail_tsn as _idt
    canon = compare_env._id_canonical_header
    cur, leg = list(_idt._TSMIS_HEADER[1:]), list(_idt._TSMIS_HEADER_LEGACY[1:])
    check("both editions canonicalize to the SAME header", canon(cur) == canon(leg))
    check("the canonical header is the current labels", canon(leg) == cur)
    check("a leading-Route consolidated header canonicalizes too",
          canon(["Route"] + leg) == ["Route"] + cur)
    check("an unrecognized header is returned UNCHANGED (identity, not refused)",
          canon(["P", "Post Mile", "S"]) == ["P", "Post Mile", "S"])

    root = Path(tempfile.mkdtemp(prefix="idedition_"))
    a = root / "2026-07-17 ssor-prod" / "intersection_detail"     # current edition
    b = root / "2026-07-08 ssor-prod" / "intersection_detail"     # legacy edition
    a.mkdir(parents=True); b.mkdir(parents=True)
    # Same physical intersection on both sides; Description differs (an aligned diff).
    _write_id_route_hdr(a, "001", cur, [_id_full_row(Description="MAIN ST")])
    _write_id_route_hdr(b, "001", leg, [_id_full_row(Description="OTHER ST")])
    out = root / "cmp.xlsx"
    res = compare_env.INTERSECTION_DETAIL.compare_folders(
        str(a.parent), str(b.parent), str(out), events=Events(),
        confirm_overwrite=lambda _p: True, mode="values",
        labels=("2026-07-17 new", "7.8 old"))
    check("cross-edition compare ok (no layout refusal)", res.status == "ok")
    if res.status == "ok":
        header, rows = _comparison_rows(out)
        di = header.index("Description")
        check("the position-aligned Description diff is flagged (MAIN ST vs OTHER ST)",
              any(DIFF in r[di] and "MAIN ST" in r[di] and "OTHER ST" in r[di]
                  for r in rows))
        check("exactly one differing cell (only Description changed)",
              res.comparison_outcome is not None
              and res.comparison_outcome.counts.differing_cells == 1)


def main():
    test_wiring()
    test_summary_aggregate_compare()
    test_detail_flat_compare()
    test_detail_edition_boundary()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-ENV-INTERSECTION CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
