"""Golden check for the cross-environment Intersection Summary adapter
(compare_env.INTERSECTION_SUMMARY) — the AGGREGATE-per-route env recipe added v0.17.0.

Locks: the adapter wiring (aggregate path — side_loader set, sheet_name None,
agg_header == IS_HEADER); its promotion to a full matrix row (in COMPARE_REPORTS as a
folders/env adapter and in reports.matrix_rows, removed from the TSN-only extra rows so
it isn't duplicated); and end-to-end that EnvCompare's aggregate path keys on Route
(has_route=False), flags a genuine per-route difference, and emits the env-labelled
discrepancy workbook. The real per-route block-walk parser is golden-tested by
check_consolidate_intersection; this exercises the env-adapter machinery via a stub
side-loader (CI-safe; no category sheets to craft).

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_env_intersection_summary.py
"""
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import compare_env
import reports
from events import Events
from openpyxl import load_workbook

_fail = []
DIFF = " ≠ "


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


def test_wiring():
    print("adapter wiring + matrix promotion:")
    a = compare_env.INTERSECTION_SUMMARY
    check("key + subdir = intersection_summary",
          a.key == "intersection_summary" and a.subdir == "intersection_summary")
    check("aggregate path: side_loader set, sheet_name None",
          a.side_loader is compare_env._load_intersection_summary_side
          and a.sheet_name is None)
    check("agg_header is IS_HEADER (Route + Total + categories)",
          a.agg_header is compare_env.IS_HEADER
          and a.agg_header[:2] == ["Route", "Total Intersections"]
          and len(a.agg_header) > 2)
    check("registered in COMPARE_REPORTS as a folders/env row",
          any(adapter is a and kind == "folders" and group == "env"
              for _l, adapter, kind, group in reports.COMPARE_REPORTS))
    check("promoted to a matrix row (cross-env)",
          "intersection_summary" in {r[0] for r in reports.matrix_rows()})
    check("removed from the TSN-only extra rows (not duplicated in the by-day matrix)",
          "intersection_summary" not in {r[0] for r in reports.tsn_matrix_extra_rows()})


def _summary(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        text = []
        for r in wb["Summary"].iter_rows(values_only=True):
            text += [str(c) for c in r if c is not None and str(c).strip()]
        return text, wb.sheetnames
    finally:
        wb.close()


def test_aggregate_compare():
    print("aggregate env compare (route-keyed; one genuine per-route diff):")
    H = compare_env.IS_HEADER
    ncat = len(H) - 2                      # category columns after Route + Total

    def mk(route, total, firstcat):
        return [route, total, firstcat] + [0] * (ncat - 1)

    # Same two routes both sides; route 002's Total differs (5 vs 6) on the B side.
    def stub(folder, label, events):       # noqa: ARG001
        is_b = "ars" in str(folder).lower()
        return [mk("001", 10, 3), mk("002", 6 if is_b else 5, 1)], []

    # A fresh EnvCompare with the real agg_header + base schema but a stub loader,
    # so the aggregate compare_folders machinery is exercised without real files.
    adapter = compare_env.EnvCompare(
        "intersection_summary", "Intersection Summary", "intersection_summary",
        side_loader=stub, agg_header=H,
        base_schema=compare_env.INTERSECTION_SUMMARY.base_schema)

    root = Path(tempfile.mkdtemp(prefix="isenv_"))
    dir_a = root / "2026-06-19 ssor-prod"
    dir_b = root / "2026-06-19 ars-prod"
    dir_a.mkdir(parents=True)
    dir_b.mkdir(parents=True)
    out = root / "cmp.xlsx"
    res = adapter.compare_folders(str(dir_a), str(dir_b), str(out), events=Events(),
                                  confirm_overwrite=lambda _p: True, mode="values")
    check("compare ok", res.status == "ok")
    text, sheets = _summary(out)
    blob = " ".join(text)
    check("side labels are the environments (SSOR-PROD vs ARS-PROD)",
          "SSOR-PROD" in blob and "ARS-PROD" in blob)
    check("keyed on Route", "keyed on Route" in blob)
    check("exactly one differing cell (route 002 Total)", "1 differing cell" in blob)
    check("both routes matched (no one-sided)", "0 one-sided" in blob or "0 row" in blob)
    check("per-environment side sheets present",
          "SSOR-PROD" in sheets and "ARS-PROD" in sheets and "Comparison" in sheets)


def main():
    test_wiring()
    test_aggregate_compare()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-ENV-INTERSECTION-SUMMARY CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
