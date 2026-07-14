#!/usr/bin/env python3
"""Independent Ramp TSN XLSX/PDF cross-format oracle.

This deliberately does not import any application parser, comparator, evidence
builder, or application constant.  Its schema, geometry, and category mapping
were measured directly from the three hash-bound source artifacts below.
"""

from __future__ import annotations

import argparse
import collections
import datetime as dt
import hashlib
import json
import re
import sys
import traceback
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import pdfplumber
from pdfminer.pdftypes import resolve1


SOURCE_BINDINGS = {
    "xlsx": {
        "sha256": "3e0c552a0a130db07275eed776a05f2a3bd0b438b53eb33ceec54bdd9c722856",
        "bytes": 1_590_431,
    },
    "detail_pdf": {
        "sha256": "0d1e31054e8f866de3be924ba350a5bd77f9230d453e58d761dea079f4505a49",
        "bytes": 1_384_895,
        "pages": 500,
        "width": 792.0,
        "height": 612.0,
    },
    "summary_pdf": {
        "sha256": "e09842e939af4bc0da82014cfd0de1f6670e7fed5e4c5f6441628bda818a118b",
        "bytes": 11_758,
        "pages": 3,
        "width": 792.0,
        "height": 612.0,
    },
}

XLSX_HEADERS = (
    "RAM_CONNECTION_ID",
    "RAMP_NANE",
    "LOCATION",
    "PR",
    "PM",
    "PM_SFX",
    "DATE_OF_RECORD",
    "HG",
    "AREA_4",
    "CITY_CODE",
    "POP",
    "ON_OFF",
    "ADT_EFF_YEAR",
    "ADT",
    "RAMP_TYPE",
    "EFF_DATE",
    "DESCRIPTION",
    "SEG_ORDER_ID",
)

# Boundaries are the midpoints between the fixed Oracle Reports columns on the
# measured 792x612 data pages.  PM_SFX has no printed column in this PDF.
DETAIL_COLUMNS = (
    ("LOCATION", 0.0, 75.0),
    ("PR", 75.0, 91.5),
    ("PM", 91.5, 140.0),
    ("DATE_OF_RECORD", 140.0, 215.0),
    ("HG", 215.0, 241.0),
    ("AREA_4", 241.0, 263.0),
    ("CITY_CODE", 263.0, 300.0),
    ("POP", 300.0, 318.0),
    ("ON_OFF", 318.0, 336.0),
    ("ADT_EFF_YEAR", 336.0, 370.0),
    ("ADT", 370.0, 420.0),
    ("RAMP_TYPE", 420.0, 444.0),
    ("EFF_DATE", 444.0, 525.0),
    ("DESCRIPTION", 525.0, 792.1),
)

DESCRIPTION_CLIP_BINDING = {
    "clip_x": 539.36,
    "clip_width": 165.32,
    "clip_right": 704.68,
    "clip_height": 12.0,
    "text_x": 539.36,
    "font": "Helvetica",
    "font_size": 8.0,
}

CLIPPED_TEXT_RUN_RE = re.compile(
    rb"q\s+"
    rb"([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s+re\s+W\s+n\s+"
    rb"BT\s+([0-9.]+)\s+([0-9.]+)\s+TD\s+"
    rb"(?:/F\d+\s+[0-9.]+\s+Tf\s+)?"
    rb"\(((?:\\.|[^\\)])*)\)\s+Tj\s+ET\s+Q",
    re.DOTALL,
)

PRINTED_FIELDS = tuple(name for name, _, _ in DETAIL_COLUMNS)
KEY_FIELDS = ("LOCATION", "PR", "PM")
SOURCE_ONLY_FIELDS = ("RAM_CONNECTION_ID", "RAMP_NANE", "SEG_ORDER_ID")
XLSX_FIELD_DISPOSITION = {
    **{field: "printed_in_detail_pdf_and_compared" for field in PRINTED_FIELDS},
    "PM_SFX": "not_printed; exact source invariant asserted against printed HG",
    **{field: "xlsx_source_only; raw_to_normalized conservation belongs to Stage 6"
       for field in SOURCE_ONLY_FIELDS},
}

# pdfminer exposes the four source newline characters in the bound PDF as
# ``(cid:13)``.  Only these exact physical records may receive that extraction-
# artifact normalization; any extra, missing, relocated, or changed token is a
# new unresolved source fact.
CID13_ARTIFACT_CONTRACT = (
    {
        "key": "08-RIV-010|R|071.863", "physical_page": 347, "report_page": 345,
        "row_on_page": 25,
        "raw_extraction": "EBOFF TO CACTUS CITY REST AREA(cid:13)",
    },
    {
        "key": "08-RIV-010|R|072.028", "physical_page": 347, "report_page": 345,
        "row_on_page": 26,
        "raw_extraction": "WB ON FR CACTUS CITY REST AREA(cid:13)",
    },
    {
        "key": "08-RIV-010|R|072.200", "physical_page": 347, "report_page": 345,
        "row_on_page": 27,
        "raw_extraction": "EB ON FR CACTUS CITY REST AREA(cid:13)",
    },
    {
        "key": "08-RIV-010|R|072.355", "physical_page": 347, "report_page": 345,
        "row_on_page": 28,
        "raw_extraction": "WBOFF TO CACTUS CITY REST AREA(cid:13)",
    },
)
LOCATION_RE = re.compile(r"^\d{2}-[A-Z]{1,3}-[0-9A-Z]{3,4}$")
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
PM_RE = re.compile(r"^\d{3}\.\d{3}$")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def render_text(value: Any) -> str:
    return " ".join(clean(value).split())


def render_date(value: Any) -> str:
    if value is None or clean(value) == "":
        return ""
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date):
        return value.strftime("%m/%d/%Y")
    raw = clean(value)
    if DATE_RE.fullmatch(raw):
        return raw
    try:
        return dt.datetime.fromisoformat(raw).strftime("%m/%d/%Y")
    except ValueError as exc:
        raise ValueError(f"unrecognized date value {raw!r}") from exc


def key_string(key: tuple[str, str, str]) -> str:
    return "|".join(key)


def canonical_digest(items: Iterable[Any]) -> str:
    digest = hashlib.sha256()
    for item in items:
        line = json.dumps(item, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
        digest.update(line.encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def serial(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return value


def decode_pdf_literal(raw: bytes) -> str:
    output = bytearray()
    index = 0
    escapes = {
        ord("n"): ord("\n"),
        ord("r"): ord("\r"),
        ord("t"): ord("\t"),
        ord("b"): ord("\b"),
        ord("f"): ord("\f"),
        ord("("): ord("("),
        ord(")"): ord(")"),
        ord("\\"): ord("\\"),
    }
    while index < len(raw):
        value = raw[index]
        if value != ord("\\"):
            output.append(value)
            index += 1
            continue
        index += 1
        if index >= len(raw):
            output.append(ord("\\"))
            break
        value = raw[index]
        if value in escapes:
            output.append(escapes[value])
            index += 1
            continue
        if value in (ord("\r"), ord("\n")):
            if value == ord("\r") and index + 1 < len(raw) and raw[index + 1] == ord("\n"):
                index += 1
            index += 1
            continue
        if ord("0") <= value <= ord("7"):
            digits = bytes([value])
            index += 1
            while index < len(raw) and len(digits) < 3 and ord("0") <= raw[index] <= ord("7"):
                digits += bytes([raw[index]])
                index += 1
            output.append(int(digits, 8) & 0xFF)
            continue
        output.append(value)
        index += 1
    return output.decode("latin-1")


def clipped_text_runs(page: Any) -> list[dict[str, Any]]:
    stream = resolve1(page.page_obj.attrs["Contents"])
    content = stream.get_data()
    runs: list[dict[str, Any]] = []
    for match in CLIPPED_TEXT_RUN_RE.finditer(content):
        clip_x, clip_y, clip_width, clip_height, text_x, text_y = (
            float(match.group(index)) for index in range(1, 7)
        )
        runs.append(
            {
                "clip_x": clip_x,
                "clip_y": clip_y,
                "clip_width": clip_width,
                "clip_height": clip_height,
                "clip_right": clip_x + clip_width,
                "text_x": text_x,
                "text_y": text_y,
                "text_top": float(page.height) - text_y - 6.4,
                "content_text": decode_pdf_literal(match.group(7)),
            }
        )
    return runs


def check_source(path: Path, binding: dict[str, Any]) -> dict[str, Any]:
    actual = {"path": str(path), "bytes": path.stat().st_size, "sha256": sha256(path)}
    actual["binding_ok"] = (
        actual["bytes"] == binding["bytes"] and actual["sha256"] == binding["sha256"]
    )
    return actual


def cid13_artifact_contract(observed: list[dict[str, Any]]) -> dict[str, Any]:
    fields = ("key", "physical_page", "report_page", "row_on_page", "raw_extraction")
    actual = [{field: item.get(field) for field in fields} for item in observed]
    expected = [{field: item[field] for field in fields} for item in CID13_ARTIFACT_CONTRACT]
    return {"pass": actual == expected, "expected": expected, "actual": actual}


def internal_negative_mutation_self_check() -> dict[str, Any]:
    exact = [dict(item) for item in CID13_ARTIFACT_CONTRACT]
    changed_identity = [dict(item) for item in CID13_ARTIFACT_CONTRACT]
    changed_identity[0]["key"] = "08-RIV-010|R|071.864"
    extra = [*exact, {**exact[-1], "row_on_page": 29}]
    field_disposition_exact = (
        set(XLSX_FIELD_DISPOSITION) == set(XLSX_HEADERS)
        and len(XLSX_FIELD_DISPOSITION) == len(XLSX_HEADERS)
    )
    checks = {
        "exact_four_artifacts_accepted": cid13_artifact_contract(exact)["pass"],
        "missing_artifact_rejected": not cid13_artifact_contract(exact[:-1])["pass"],
        "extra_artifact_rejected": not cid13_artifact_contract(extra)["pass"],
        "changed_identity_rejected": not cid13_artifact_contract(changed_identity)["pass"],
        "all_18_xlsx_fields_have_one_disposition": field_disposition_exact,
    }
    return {"checks": checks, "pass": all(checks.values())}


def load_xlsx(path: Path, unresolved: list[dict[str, Any]]) -> tuple[dict[Any, Any], dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    metadata = {
        "sheet_names": workbook.sheetnames,
        "created": serial(workbook.properties.created),
        "modified": serial(workbook.properties.modified),
    }
    if len(workbook.sheetnames) != 1:
        unresolved.append({"kind": "xlsx_sheet_count", "actual": workbook.sheetnames})
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = tuple(clean(value) for value in next(iterator))
    metadata.update({"headers": list(headers), "max_row": sheet.max_row, "max_column": sheet.max_column})
    if headers != XLSX_HEADERS:
        unresolved.append(
            {"kind": "xlsx_schema", "expected": list(XLSX_HEADERS), "actual": list(headers)}
        )

    records: dict[tuple[str, str, str], dict[str, str]] = {}
    raw_rows: list[tuple[Any, ...]] = []
    duplicate_keys: dict[str, int] = collections.Counter()
    for excel_row, values in enumerate(iterator, 2):
        row = tuple(values)
        if len(row) != len(XLSX_HEADERS) or all(value is None for value in row):
            unresolved.append({"kind": "xlsx_row_shape", "excel_row": excel_row, "values": row})
            continue
        raw_rows.append(row)
        record = {
            "LOCATION": clean(row[2]),
            "PR": clean(row[3]),
            "PM": clean(row[4]),
            "DATE_OF_RECORD": render_date(row[6]),
            "HG": clean(row[7]),
            "AREA_4": clean(row[8]),
            "CITY_CODE": clean(row[9]),
            "POP": clean(row[10]),
            "ON_OFF": clean(row[11]),
            "ADT_EFF_YEAR": clean(row[12]),
            "ADT": clean(row[13]),
            "RAMP_TYPE": clean(row[14]),
            "EFF_DATE": render_date(row[15]),
            "DESCRIPTION": render_text(row[16]),
        }
        key = tuple(record[field] for field in KEY_FIELDS)
        duplicate_keys[key_string(key)] += 1
        if key not in records:
            records[key] = {
                "excel_row": excel_row,
                "ram_connection_id": serial(row[0]),
                "ramp_nane": clean(row[1]),
                "pm_sfx": clean(row[5]),
                "seg_order_id": serial(row[17]),
                "description_source_text": clean(row[16]),
                "printed": record,
            }
    duplicate_keys = {key: count for key, count in duplicate_keys.items() if count != 1}
    if duplicate_keys:
        unresolved.append({"kind": "xlsx_ambiguous_identity", "keys": duplicate_keys})

    allowed = {
        "PR": {"", "C", "L", "M", "R", "S", "T"},
        "PM_SFX": {"", "L", "R"},
        "HG": {"D", "L", "R", "U"},
        "AREA_4": {"N", "Y"},
        "POP": {"B", "R", "U"},
        "ON_OFF": {"F", "O", "Z"},
        "ADT_EFF_YEAR": {"2023"},
        "RAMP_TYPE": set("ABCDEFGHJKLMPRVZ"),
    }
    column_index = {name: index for index, name in enumerate(XLSX_HEADERS)}
    unexpected_domains: dict[str, list[str]] = {}
    for field, expected in allowed.items():
        actual = {clean(row[column_index[field]]) for row in raw_rows}
        if not actual <= expected:
            unexpected_domains[field] = sorted(actual - expected)
    if unexpected_domains:
        unresolved.append({"kind": "xlsx_unexpected_domain", "fields": unexpected_domains})

    suffix_counts = collections.Counter(clean(row[5]) for row in raw_rows if clean(row[5]))
    suffix_hg_disagreement = sum(
        clean(row[5]) != clean(row[7]) for row in raw_rows if clean(row[5])
    )
    blank_suffix_hg_lr = sum(
        not clean(row[5]) and clean(row[7]) in {"L", "R"} for row in raw_rows
    )
    suffix_rule = {
        "nonblank": sum(suffix_counts.values()),
        "by_value": dict(sorted(suffix_counts.items())),
        "nonblank_not_equal_hg": suffix_hg_disagreement,
        "blank_suffix_with_hg_l_or_r": blank_suffix_hg_lr,
        "expected": {"nonblank": 313, "by_value": {"L": 165, "R": 148}},
    }
    suffix_rule["ok"] = (
        suffix_rule["nonblank"] == 313
        and suffix_rule["by_value"] == {"L": 165, "R": 148}
        and suffix_hg_disagreement == 0
        and blank_suffix_hg_lr == 0
    )
    if not suffix_rule["ok"]:
        unresolved.append({"kind": "pm_sfx_hg_rule", "details": suffix_rule})

    source_only_census: dict[str, dict[str, Any]] = {}
    for field in SOURCE_ONLY_FIELDS:
        values = [serial(row[column_index[field]]) for row in raw_rows]
        source_only_census[field] = {
            "rows": len(values),
            "nonblank": sum(clean(value) != "" for value in values),
            "unique_typed_values": len({
                json.dumps(value, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
                for value in values
            }),
            "typed_values_digest": canonical_digest(values),
            "disposition": XLSX_FIELD_DISPOSITION[field],
        }

    field_disposition = {
        "columns": [
            {"xlsx_column": field, "disposition": XLSX_FIELD_DISPOSITION[field]}
            for field in XLSX_HEADERS
        ],
        "printed_and_compared": list(PRINTED_FIELDS),
        "relationally_asserted_not_printed": ["PM_SFX"],
        "xlsx_source_only": list(SOURCE_ONLY_FIELDS),
        "all_18_columns_accounted_once": (
            set(XLSX_FIELD_DISPOSITION) == set(XLSX_HEADERS)
            and len(XLSX_FIELD_DISPOSITION) == len(XLSX_HEADERS)
        ),
        "stage_boundary": (
            "This oracle classifies PDF disposition and binds source-only values; "
            "their raw-to-normalized conservation is promoted in Stage 6."
        ),
    }
    if not field_disposition["all_18_columns_accounted_once"]:
        unresolved.append({"kind": "xlsx_field_disposition_contract", "details": field_disposition})

    metadata.update(
        {
            "data_rows": len(raw_rows),
            "unique_physical_keys": len(records),
            "duplicate_physical_keys": duplicate_keys,
            "pm_sfx_rule": suffix_rule,
            "field_disposition": field_disposition,
            "source_only_field_census": source_only_census,
            "records_digest": canonical_digest(
                {"key": key, **records[key]["printed"]} for key in sorted(records)
            ),
        }
    )
    workbook.close()
    return records, {"metadata": metadata, "raw_rows": raw_rows}


def words_in_band(words: list[dict[str, Any]], top: float) -> list[dict[str, Any]]:
    return [word for word in words if abs(float(word["top"]) - top) <= 1.5]


def parse_detail_pdf(
    path: Path, unresolved: list[dict[str, Any]]
) -> tuple[dict[Any, Any], dict[str, Any]]:
    records: dict[tuple[str, str, str], dict[str, Any]] = {}
    page_rows: list[dict[str, Any]] = []
    extraction_artifacts: list[dict[str, Any]] = []
    description_clip_geometry_counts: collections.Counter[tuple[float, ...]] = collections.Counter()
    duplicate_keys: dict[str, int] = collections.Counter()
    with pdfplumber.open(path) as pdf:
        metadata = dict(pdf.metadata or {})
        dimensions = [(float(page.width), float(page.height)) for page in pdf.pages]
        bad_dimensions = [
            index + 1
            for index, (width, height) in enumerate(dimensions)
            if abs(width - 792.0) > 0.01 or abs(height - 612.0) > 0.01
        ]
        if len(pdf.pages) != SOURCE_BINDINGS["detail_pdf"]["pages"]:
            unresolved.append({"kind": "detail_page_count", "actual": len(pdf.pages)})
        if bad_dimensions:
            unresolved.append({"kind": "detail_page_geometry", "pages": bad_dimensions})

        cover = pdf.pages[0].extract_text(x_tolerance=1, y_tolerance=2) or ""
        parameters = pdf.pages[1].extract_text(x_tolerance=1, y_tolerance=2) or ""
        for required in ("OTM22260", "TSAR - RAMPS DETAIL"):
            if required not in cover:
                unresolved.append({"kind": "detail_cover_marker", "missing": required})
        parameter_patterns = {
            "report_date": r"REPORT DATE\s*:\s*(\d{2}/\d{2}/\d{4})",
            "reference_date": r"REFERENCE DATE\s*:\s*(\d{2}/\d{2}/\d{4})",
            "event_id": r"EVENT ID\s*:\s*(\d+)",
        }
        parameters_parsed: dict[str, str] = {}
        for field, pattern in parameter_patterns.items():
            match = re.search(pattern, parameters)
            if match:
                parameters_parsed[field] = match.group(1)
            else:
                unresolved.append({"kind": "detail_parameter", "field": field})

        for physical_page, page in enumerate(pdf.pages[2:], 3):
            report_page = physical_page - 2
            words = page.extract_words(x_tolerance=1, y_tolerance=1, keep_blank_chars=False)
            page_clip_runs = clipped_text_runs(page)
            header_text = page.extract_text(x_tolerance=1, y_tolerance=2) or ""
            page_match = re.search(r"Page#\s+(\d+)", header_text)
            actual_report_page = int(page_match.group(1)) if page_match else None
            if actual_report_page != report_page:
                unresolved.append(
                    {
                        "kind": "detail_report_page_sequence",
                        "physical_page": physical_page,
                        "expected": report_page,
                        "actual": actual_report_page,
                    }
                )

            anchors = [
                word
                for word in words
                if float(word["x0"]) < 70.0 and LOCATION_RE.fullmatch(word["text"])
            ]
            anchors.sort(key=lambda word: float(word["top"]))
            page_rows.append(
                {
                    "physical_page": physical_page,
                    "report_page": report_page,
                    "rows": len(anchors),
                }
            )
            for row_number, anchor in enumerate(anchors, 1):
                band = words_in_band(words, float(anchor["top"]))
                record: dict[str, str] = {}
                assigned: set[int] = set()
                for field, low, high in DETAIL_COLUMNS:
                    selected = [
                        (index, word)
                        for index, word in enumerate(band)
                        if low <= float(word["x0"]) < high
                    ]
                    selected.sort(key=lambda item: float(item[1]["x0"]))
                    assigned.update(index for index, _ in selected)
                    record[field] = render_text(" ".join(word["text"] for _, word in selected))
                key = tuple(record[field] for field in KEY_FIELDS)
                if "(cid:13)" in record["DESCRIPTION"]:
                    artifact = {
                        "key": key_string(key),
                        "physical_page": physical_page,
                        "report_page": report_page,
                        "row_on_page": row_number,
                        "raw_extraction": record["DESCRIPTION"],
                    }
                    contracted = artifact in CID13_ARTIFACT_CONTRACT
                    extraction_artifacts.append({
                        **artifact,
                        "contracted": contracted,
                        "classification": (
                            "exact hash/page/row/identity-bound source newline extraction artifact"
                            if contracted else "unexpected cid:13 extraction token; not normalized"
                        ),
                    })
                    if contracted:
                        record["DESCRIPTION"] = render_text(
                            record["DESCRIPTION"].replace("(cid:13)", "")
                        )
                    else:
                        unresolved.append({"kind": "unexpected_pdf_cid13_artifact", **artifact})
                description_chars = [
                    char
                    for char in page.chars
                    if 525.0 <= float(char["x0"]) < 792.1
                    and abs(float(char["top"]) - float(anchor["top"])) <= 1.5
                ]
                description_runs = [
                    run
                    for run in page_clip_runs
                    if abs(float(run["text_top"]) - float(anchor["top"])) <= 2.0
                    and render_text(run["content_text"].replace("\r", ""))
                    == record["DESCRIPTION"]
                ]
                description_geometry = None
                if len(description_runs) == 1 and description_chars:
                    run = description_runs[0]
                    fonts = sorted({clean(char.get("fontname")) for char in description_chars})
                    font_sizes = sorted({round(float(char.get("size", 0.0)), 3) for char in description_chars})
                    description_geometry = {
                        **run,
                        "rendered_text_start_x": min(float(char["x0"]) for char in description_chars),
                        "rendered_text_end_x": max(float(char["x1"]) for char in description_chars),
                        "fonts": fonts,
                        "font_sizes": font_sizes,
                    }
                    geometry_key = (
                        round(float(run["clip_x"]), 3),
                        round(float(run["clip_width"]), 3),
                        round(float(run["clip_right"]), 3),
                        round(float(run["clip_height"]), 3),
                        round(float(run["text_x"]), 3),
                    )
                    description_clip_geometry_counts[geometry_key] += 1
                elif len(description_runs) > 1:
                    unresolved.append(
                        {
                            "kind": "detail_ambiguous_description_clip_run",
                            "physical_page": physical_page,
                            "row": row_number,
                            "description": record["DESCRIPTION"],
                            "runs": description_runs,
                        }
                    )
                unassigned = [band[index]["text"] for index in range(len(band)) if index not in assigned]
                if unassigned:
                    unresolved.append(
                        {
                            "kind": "detail_unassigned_row_words",
                            "physical_page": physical_page,
                            "row": row_number,
                            "words": unassigned,
                        }
                    )

                required_fields = set(PRINTED_FIELDS) - {"PR", "CITY_CODE"}
                missing = sorted(field for field in required_fields if not record[field])
                pattern_failures: dict[str, str] = {}
                for field, pattern in (
                    ("LOCATION", LOCATION_RE),
                    ("PM", PM_RE),
                    ("DATE_OF_RECORD", DATE_RE),
                    ("EFF_DATE", DATE_RE),
                ):
                    if record[field] and not pattern.fullmatch(record[field]):
                        pattern_failures[field] = record[field]
                for field, pattern in (
                    ("HG", r"[DLRU]"),
                    ("AREA_4", r"[NY]"),
                    ("POP", r"[BRU]"),
                    ("ON_OFF", r"[FOZ]"),
                    ("ADT_EFF_YEAR", r"\d{4}"),
                    ("ADT", r"\d{6}"),
                    ("RAMP_TYPE", r"[A-HJKLMPRVZ]"),
                ):
                    if record[field] and not re.fullmatch(pattern, record[field]):
                        pattern_failures[field] = record[field]
                if missing or pattern_failures:
                    unresolved.append(
                        {
                            "kind": "detail_row_shape",
                            "physical_page": physical_page,
                            "row": row_number,
                            "missing": missing,
                            "pattern_failures": pattern_failures,
                            "record": record,
                        }
                    )

                duplicate_keys[key_string(key)] += 1
                if key not in records:
                    records[key] = {
                        "physical_page": physical_page,
                        "report_page": report_page,
                        "row_on_page": row_number,
                        "description_print_geometry": description_geometry,
                        "printed": record,
                    }

        duplicate_keys = {key: count for key, count in duplicate_keys.items() if count != 1}
        if duplicate_keys:
            unresolved.append({"kind": "detail_ambiguous_identity", "keys": duplicate_keys})
        if len(records) != sum(item["rows"] for item in page_rows):
            unresolved.append(
                {
                    "kind": "detail_duplicate_or_lost_rows",
                    "unique": len(records),
                    "parsed": sum(item["rows"] for item in page_rows),
                }
            )
        artifact_contract = cid13_artifact_contract(extraction_artifacts)
        if not artifact_contract["pass"]:
            unresolved.append({
                "kind": "pdf_cid13_artifact_contract_failed",
                "details": artifact_contract,
            })
        return records, {
            "pdf_metadata": metadata,
            "parameters": parameters_parsed,
            "physical_pages": len(dimensions),
            "data_pages": len(page_rows),
            "parsed_rows": sum(item["rows"] for item in page_rows),
            "unique_physical_keys": len(records),
            "duplicate_physical_keys": duplicate_keys,
            "pdf_text_extraction_artifacts": {
                "count": len(extraction_artifacts),
                "items": extraction_artifacts,
                "contract": artifact_contract,
            },
            "description_clip_geometry_counts": [
                {
                    "clip_x": key[0],
                    "clip_width": key[1],
                    "clip_right": key[2],
                    "clip_height": key[3],
                    "text_x": key[4],
                    "rows": count,
                }
                for key, count in sorted(description_clip_geometry_counts.items())
            ],
            "page_rows": page_rows,
            "records_digest": canonical_digest(
                {"key": key, **records[key]["printed"]} for key in sorted(records)
            ),
        }


def nearest_row(words: list[dict[str, Any]], count_word: dict[str, Any], code_x: tuple[float, float]) -> str:
    candidates = [
        word
        for word in words
        if code_x[0] <= float(word["x0"]) < code_x[1]
        and abs(float(word["top"]) - float(count_word["top"])) <= 1.5
    ]
    candidates.sort(key=lambda word: float(word["x0"]))
    return render_text(" ".join(word["text"] for word in candidates))


def parse_summary_pdf(path: Path, unresolved: list[dict[str, Any]]) -> dict[str, Any]:
    with pdfplumber.open(path) as pdf:
        dimensions = [(float(page.width), float(page.height)) for page in pdf.pages]
        if len(pdf.pages) != SOURCE_BINDINGS["summary_pdf"]["pages"]:
            unresolved.append({"kind": "summary_page_count", "actual": len(pdf.pages)})
        bad_dimensions = [
            index + 1
            for index, (width, height) in enumerate(dimensions)
            if abs(width - 792.0) > 0.01 or abs(height - 612.0) > 0.01
        ]
        if bad_dimensions:
            unresolved.append({"kind": "summary_page_geometry", "pages": bad_dimensions})
        cover = pdf.pages[0].extract_text(x_tolerance=1, y_tolerance=2) or ""
        parameters = pdf.pages[1].extract_text(x_tolerance=1, y_tolerance=2) or ""
        page = pdf.pages[2]
        words = page.extract_words(x_tolerance=1, y_tolerance=1, keep_blank_chars=False)
        for required in ("OTM22270", "TSAR - RAMPS SUMMARY"):
            if required not in cover:
                unresolved.append({"kind": "summary_cover_marker", "missing": required})

        parameter_patterns = {
            "report_date": r"REPORT DATE\s*:\s*(\d{2}/\d{2}/\d{4})",
            "reference_date": r"REFERENCE DATE\s*:\s*(\d{2}/\d{2}/\d{4})",
            "event_id": r"EVENT ID\s*:\s*(\d+)",
        }
        parameters_parsed: dict[str, str] = {}
        for field, pattern in parameter_patterns.items():
            match = re.search(pattern, parameters)
            if match:
                parameters_parsed[field] = match.group(1)
            else:
                unresolved.append({"kind": "summary_parameter", "field": field})

        highway: dict[str, int] = {}
        for count_word in words:
            x0, top, token = float(count_word["x0"]), float(count_word["top"]), count_word["text"]
            if not (70.0 <= x0 < 120.0 and 115.0 <= top <= 215.0 and token.isdigit()):
                continue
            code = nearest_row(words, count_word, (145.0, 230.0)).split(" - ")[0].split()[0]
            highway[code] = int(token)

        on_off: dict[str, int] = {}
        for count_word in words:
            x0, top, token = float(count_word["x0"]), float(count_word["top"]), count_word["text"]
            if not (70.0 <= x0 < 120.0 and 280.0 <= top <= 330.0 and token.isdigit()):
                continue
            code = nearest_row(words, count_word, (145.0, 225.0)).split()[0]
            on_off[code] = int(token)

        population_words = sorted(
            (
                word
                for word in words
                if 70.0 <= float(word["x0"]) < 120.0
                and 380.0 <= float(word["top"]) <= 475.0
                and word["text"].isdigit()
            ),
            key=lambda word: float(word["top"]),
        )
        population_labels = ("R-I", "R-O", "U-I", "U-O", "INVALID")
        population = {
            label: int(word["text"])
            for label, word in zip(population_labels, population_words, strict=False)
        }
        if len(population_words) != len(population_labels):
            unresolved.append(
                {
                    "kind": "summary_population_rows",
                    "actual": [(word["text"], word["top"]) for word in population_words],
                }
            )

        ramp_types: dict[str, int] = {}
        for count_word in words:
            x0, top, token = float(count_word["x0"]), float(count_word["top"]), count_word["text"]
            if not (370.0 <= x0 < 410.0 and 115.0 <= top <= 405.0 and token.isdigit()):
                continue
            code = nearest_row(words, count_word, (475.0, 505.0)).split()[0]
            ramp_types[code] = int(token)

        total_words = [
            word
            for word in words
            if 495.0 <= float(word["x0"]) < 540.0
            and 450.0 <= float(word["top"]) <= 470.0
            and word["text"].isdigit()
        ]
        total = int(total_words[0]["text"]) if len(total_words) == 1 else None
        if total is None:
            unresolved.append(
                {"kind": "summary_total", "actual": [(word["text"], word["top"]) for word in total_words]}
            )

        expected_category_shapes = {
            "highway_groups": {"R", "D", "U", "X", "L", "Others"},
            "on_off": {"ON", "OFF", "OTH"},
            "population_groups": set(population_labels),
            "ramp_types": set("ABCDEFGHJKLMPRVZ"),
        }
        actual_shapes = {
            "highway_groups": set(highway),
            "on_off": set(on_off),
            "population_groups": set(population),
            "ramp_types": set(ramp_types),
        }
        for section, expected in expected_category_shapes.items():
            if actual_shapes[section] != expected:
                unresolved.append(
                    {
                        "kind": "summary_category_shape",
                        "section": section,
                        "expected": sorted(expected),
                        "actual": sorted(actual_shapes[section]),
                    }
                )

        return {
            "pdf_metadata": dict(pdf.metadata or {}),
            "parameters": parameters_parsed,
            "physical_pages": len(pdf.pages),
            "categories": {
                "highway_groups": highway,
                "on_off": on_off,
                "population_groups": population,
                "ramp_types": ramp_types,
            },
            "total": total,
        }


def aggregate_xlsx(raw_rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    highway_raw = collections.Counter(clean(row[7]) for row in raw_rows)
    highway = {
        "R": highway_raw["R"],
        "D": highway_raw["D"],
        "U": highway_raw["U"],
        "X": highway_raw["X"],
        "L": highway_raw["L"],
        "Others": sum(
            count for code, count in highway_raw.items() if code not in {"R", "D", "U", "X", "L"}
        ),
    }
    on_off_raw = collections.Counter(clean(row[11]) for row in raw_rows)
    on_off = {
        "ON": on_off_raw["O"],
        "OFF": on_off_raw["F"],
        "OTH": sum(count for code, count in on_off_raw.items() if code not in {"O", "F"}),
    }
    population = collections.Counter()
    for row in raw_rows:
        pop = clean(row[10])
        city_present = bool(clean(row[9]))
        if pop == "R":
            population["R-I" if city_present else "R-O"] += 1
        elif pop in {"U", "B"}:
            population["U-I" if city_present else "U-O"] += 1
        else:
            population["INVALID"] += 1
    for category in ("R-I", "R-O", "U-I", "U-O", "INVALID"):
        population[category] += 0
    ramp_types = collections.Counter(clean(row[14]) for row in raw_rows)
    return {
        "highway_groups": highway,
        "on_off": on_off,
        "population_groups": dict(population),
        "ramp_types": dict(sorted(ramp_types.items())),
        "total": len(raw_rows),
        "row_to_category_mapping": {
            "highway_groups": "HG: R/D/U/X/L; all other values -> Others",
            "on_off": "ON_OFF: O -> ON, F -> OFF, all other values -> OTH",
            "population_groups": (
                "POP R -> RURAL; POP U or B -> URBAN; CITY_CODE present -> INSIDE CITY; "
                "CITY_CODE blank -> OUTSIDE CITY; other POP -> INVALID"
            ),
            "ramp_types": "RAMP_TYPE code counted directly",
        },
    }


def compare_records(
    xlsx: dict[Any, Any], pdf: dict[Any, Any], cross_snapshot: bool, unresolved: list[dict[str, Any]]
) -> dict[str, Any]:
    xlsx_keys = set(xlsx)
    pdf_keys = set(pdf)
    raw_only_keys = sorted(xlsx_keys - pdf_keys)
    pdf_only_keys = sorted(pdf_keys - xlsx_keys)
    extracted_content_exact = 0
    field_mismatch_counts: collections.Counter[str] = collections.Counter()
    print_truncation_equivalences: list[dict[str, Any]] = []
    unproven_differences: list[dict[str, Any]] = []
    for key in sorted(xlsx_keys & pdf_keys):
        raw_record = xlsx[key]["printed"]
        pdf_record = pdf[key]["printed"]
        differences = {
            field: {"xlsx": raw_record[field], "pdf": pdf_record[field]}
            for field in PRINTED_FIELDS
            if raw_record[field] != pdf_record[field]
        }
        if not differences:
            extracted_content_exact += 1
            continue
        field_mismatch_counts.update(differences.keys())
        geometry = pdf[key].get("description_print_geometry")
        source_text = xlsx[key].get("description_source_text", "")
        content_text = geometry.get("content_text", "") if geometry else ""
        binding_ok = bool(geometry) and all(
            abs(float(geometry[field]) - float(DESCRIPTION_CLIP_BINDING[field])) <= 0.01
            for field in ("clip_x", "clip_width", "clip_right", "clip_height", "text_x")
        )
        font_ok = bool(geometry) and geometry.get("fonts") == [
            DESCRIPTION_CLIP_BINDING["font"]
        ] and geometry.get("font_sizes") == [DESCRIPTION_CLIP_BINDING["font_size"]]
        exact_prefix = (
            bool(content_text)
            and source_text.startswith(content_text)
            and len(source_text) > len(content_text)
            and render_text(content_text) == pdf_record["DESCRIPTION"]
        )
        last_pdf_content_character_end_x = (
            float(geometry["rendered_text_end_x"]) if geometry is not None else None
        )
        pdf_content_reaches_past_clip = (
            last_pdf_content_character_end_x is not None
            and last_pdf_content_character_end_x
            >= float(DESCRIPTION_CLIP_BINDING["clip_right"]) - 0.01
        )
        if (
            set(differences) == {"DESCRIPTION"}
            and exact_prefix
            and binding_ok
            and font_ok
            and pdf_content_reaches_past_clip
        ):
            print_truncation_equivalences.append(
                {
                    "key": key_string(key),
                    "excel_row": xlsx[key]["excel_row"],
                    "pdf_physical_page": pdf[key]["physical_page"],
                    "pdf_row_on_page": pdf[key]["row_on_page"],
                    "xlsx_description": source_text,
                    "pdf_content_description": content_text,
                    "omitted_suffix": source_text[len(content_text) :],
                    "proof": {
                        "exact_content_prefix": True,
                        "clip_x": geometry["clip_x"],
                        "clip_width": geometry["clip_width"],
                        "clip_right": geometry["clip_right"],
                        "text_x": geometry["text_x"],
                        "font": geometry["fonts"][0],
                        "font_size": geometry["font_sizes"][0],
                        "last_pdf_content_character_end_x": last_pdf_content_character_end_x,
                        "last_pdf_content_character_end_outside_clip_by": (
                            last_pdf_content_character_end_x - float(geometry["clip_right"])
                        ),
                        "claim": (
                            "The PDF content is an exact prefix of the XLSX source and its final "
                            "content character already ends beyond the fixed visible clip. This "
                            "does not claim to measure a character absent from the PDF stream."
                        ),
                    },
                }
            )
            continue
        unproven_differences.append(
            {
                "key": key_string(key),
                "excel_row": xlsx[key]["excel_row"],
                "pdf_physical_page": pdf[key]["physical_page"],
                "pdf_row_on_page": pdf[key]["row_on_page"],
                "differences": differences,
                "print_truncation_tests": {
                    "description_only": set(differences) == {"DESCRIPTION"},
                    "exact_content_prefix": exact_prefix,
                    "fixed_clip_binding": binding_ok,
                    "font_binding": font_ok,
                    "last_pdf_content_character_end_reaches_past_clip": (
                        pdf_content_reaches_past_clip
                    ),
                },
            }
        )

    raw_only = [
        {
            "key": key_string(key),
            "excel_row": xlsx[key]["excel_row"],
            "record": xlsx[key]["printed"],
        }
        for key in raw_only_keys
    ]
    pdf_only = [
        {
            "key": key_string(key),
            "pdf_physical_page": pdf[key]["physical_page"],
            "pdf_row_on_page": pdf[key]["row_on_page"],
            "record": pdf[key]["printed"],
        }
        for key in pdf_only_keys
    ]
    if raw_only or pdf_only or unproven_differences:
        unresolved.append(
            {
                "kind": "unproven_cross_format_difference",
                "rule": (
                    "Source dates never classify a value difference by themselves. Only DESCRIPTION "
                    "exact-prefix differences with the hash-bound fixed clip, bound font, and final "
                    "PDF content character ending past the visible clip are print-truncation equivalent."
                ),
                "raw_only": raw_only,
                "pdf_only": pdf_only,
                "matched_differences": unproven_differences,
            }
        )
    visible_exact = extracted_content_exact + len(print_truncation_equivalences)
    return {
        "exact_render_equivalence": visible_exact,
        "exact_extracted_content": extracted_content_exact,
        "proven_print_truncation_equivalence_count": len(print_truncation_equivalences),
        "proven_print_truncation_equivalences": print_truncation_equivalences,
        "source_date_export_delta": {
            "classification_basis": (
                "No difference is classified from dates alone. The later XLSX creation date is context, "
                "not evidence that an arbitrary field difference is a source-date/export delta."
            ),
            "raw_only_count": 0,
            "pdf_only_count": 0,
            "matched_record_delta_count": 0,
            "field_mismatch_counts": {},
            "raw_only": [],
            "pdf_only": [],
            "matched_record_deltas": [],
        },
        "observed_pre_render_field_differences": dict(sorted(field_mismatch_counts.items())),
        "unproven_differences": unproven_differences,
        "matched_keys": len(xlsx_keys & pdf_keys),
        "xlsx_keys": len(xlsx_keys),
        "pdf_keys": len(pdf_keys),
    }


def run(args: argparse.Namespace) -> tuple[dict[str, Any], int]:
    paths = {
        "xlsx": Path(args.xlsx).resolve(),
        "detail_pdf": Path(args.detail_pdf).resolve(),
        "summary_pdf": Path(args.summary_pdf).resolve(),
    }
    unresolved: list[dict[str, Any]] = []
    negative_check = internal_negative_mutation_self_check()
    if not negative_check["pass"]:
        unresolved.append({"kind": "internal_negative_mutation_self_check", "details": negative_check})
    sources = {
        name: check_source(path, SOURCE_BINDINGS[name]) for name, path in paths.items()
    }
    for name, source in sources.items():
        if not source["binding_ok"]:
            unresolved.append(
                {
                    "kind": "source_binding",
                    "source": name,
                    "expected": SOURCE_BINDINGS[name],
                    "actual": source,
                }
            )

    xlsx_records, xlsx_details = load_xlsx(paths["xlsx"], unresolved)
    detail_records, detail = parse_detail_pdf(paths["detail_pdf"], unresolved)
    summary = parse_summary_pdf(paths["summary_pdf"], unresolved)

    xlsx_created_raw = xlsx_details["metadata"].get("created")
    detail_reference_raw = detail.get("parameters", {}).get("reference_date")
    cross_snapshot = False
    if xlsx_created_raw and detail_reference_raw:
        try:
            xlsx_created = dt.datetime.fromisoformat(str(xlsx_created_raw)).date()
            detail_reference = dt.datetime.strptime(detail_reference_raw, "%m/%d/%Y").date()
            cross_snapshot = xlsx_created > detail_reference
        except ValueError:
            unresolved.append(
                {
                    "kind": "source_date_parse",
                    "xlsx_created": xlsx_created_raw,
                    "detail_reference": detail_reference_raw,
                }
            )

    detail_comparison = compare_records(
        xlsx_records, detail_records, cross_snapshot, unresolved
    )
    xlsx_summary = aggregate_xlsx(xlsx_details["raw_rows"])
    summary_mismatches: dict[str, Any] = {}
    for section in ("highway_groups", "on_off", "population_groups", "ramp_types"):
        if xlsx_summary[section] != summary["categories"][section]:
            summary_mismatches[section] = {
                "xlsx": xlsx_summary[section],
                "pdf": summary["categories"][section],
            }
    if xlsx_summary["total"] != summary["total"]:
        summary_mismatches["total"] = {
            "xlsx": xlsx_summary["total"],
            "pdf": summary["total"],
        }
    if summary_mismatches:
        unresolved.append(
            {
                "kind": "summary_unproven_difference",
                "rule": "Source dates alone do not classify category-count differences.",
                "details": summary_mismatches,
            }
        )

    result = {
        "oracle": "phase4_ramp_tsn_pdf_oracle",
        "oracle_version": 3,
        "independence": (
            "No production parser, comparator, evidence builder, or application constant imported."
        ),
        "sources": sources,
        "source_dates": {
            "xlsx_workbook_created": xlsx_created_raw,
            "xlsx_workbook_modified": xlsx_details["metadata"].get("modified"),
            "detail_pdf_report_date": detail.get("parameters", {}).get("report_date"),
            "detail_pdf_reference_date": detail_reference_raw,
            "summary_pdf_report_date": summary.get("parameters", {}).get("report_date"),
            "summary_pdf_reference_date": summary.get("parameters", {}).get("reference_date"),
            "cross_snapshot": cross_snapshot,
        },
        "xlsx": xlsx_details["metadata"],
        "xlsx_field_disposition": xlsx_details["metadata"]["field_disposition"],
        "internal_negative_mutation_self_check": negative_check,
        "detail_pdf": detail,
        "detail_cross_format": detail_comparison,
        "summary_pdf": summary,
        "xlsx_summary_aggregation": xlsx_summary,
        "summary_cross_format": {
            "exact": not summary_mismatches,
            "source_date_export_delta": {},
        },
        "unresolved_residue": unresolved,
    }
    source_bindings_ok = all(source["binding_ok"] for source in sources.values())
    detail_delta = detail_comparison["source_date_export_delta"]
    delta_count = (
        detail_delta["raw_only_count"]
        + detail_delta["pdf_only_count"]
        + detail_delta["matched_record_delta_count"]
    )
    result["classification"] = {
        "exact_render_equivalence_records": detail_comparison["exact_render_equivalence"],
        "exact_extracted_content_records": detail_comparison["exact_extracted_content"],
        "proven_print_truncation_equivalence_records": detail_comparison[
            "proven_print_truncation_equivalence_count"
        ],
        "source_date_export_delta_records_or_sections": delta_count,
        "unresolved_residue_count": len(unresolved),
    }
    result["status"] = "ok" if source_bindings_ok and not unresolved else "failed"
    result["parity"] = (
        "exact_with_proven_print_truncation"
        if result["status"] == "ok"
        and detail_comparison["proven_print_truncation_equivalence_count"] > 0
        and delta_count == 0
        else "exact"
        if result["status"] == "ok" and delta_count == 0
        else "classified_source_date_export_delta"
        if result["status"] == "ok"
        else "unresolved"
    )
    return result, 0 if result["status"] == "ok" else 1


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--detail-pdf", required=True)
    parser.add_argument("--summary-pdf", required=True)
    parser.add_argument("--result", required=True)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    result_path = Path(args.result).resolve()
    result_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        result, exit_code = run(args)
    except Exception as exc:  # a crash is itself fail-closed and leaves a result
        result = {
            "oracle": "phase4_ramp_tsn_pdf_oracle",
            "oracle_version": 3,
            "status": "failed",
            "parity": "unresolved",
            "unresolved_residue": [
                {
                    "kind": "oracle_exception",
                    "type": type(exc).__name__,
                    "message": str(exc),
                    "traceback": traceback.format_exc(),
                }
            ],
        }
        exit_code = 1
    temporary = result_path.with_suffix(result_path.suffix + ".tmp")
    temporary.write_text(json.dumps(result, indent=2, ensure_ascii=False, default=serial) + "\n", encoding="utf-8")
    temporary.replace(result_path)
    print(json.dumps({
        "status": result.get("status"),
        "parity": result.get("parity"),
        "classification": result.get("classification"),
        "result": str(result_path),
    }, indent=2))
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
