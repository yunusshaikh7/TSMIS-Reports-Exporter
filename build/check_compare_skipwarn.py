"""Golden check for COMPARE-SKIPPED-FILES-MATCH + CONSOLIDATE-XLSX-PARTIAL-OK.

Before: files unreadable on one/both sides (or in a consolidation) were only
logged — the comparison could still report "✓ EVERYTHING MATCHES" and the
consolidator still returned status="ok", silently dropping whole routes.

After:
  * run_compare(..., warnings=[...]) can NEVER certify a clean match: the
    verdict is forced to "diff", the summary leads with a ⚠ warning listing the
    skipped files, and the workbook banner says "COULD NOT COMPARE EVERYTHING".
    With warnings empty it is byte-identical to before (proven separately by the
    cell-dump regression check).
  * consolidate_xlsx surfaces skipped/failed inputs as a loud ⚠ INCOMPLETE
    summary line, and returns status="error" WITHOUT overwriting when nothing
    at all could be combined.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_skipwarn.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from compare_core import CompareSchema, run_compare
from consolidate_xlsx_base import consolidate_xlsx
from openpyxl import Workbook, load_workbook
from pathlib import Path

SC = CompareSchema(report_name="SkipTest", header=["Loc", "V"],
                   side_a="AENV", side_b="BENV", id_noun="row",
                   id_noun_plural="rows", sides_noun="environments")
# Two IDENTICAL sides — absent the warning this is a clean match.
ROWS = [["1.0", "x"], ["2.0", "y"], ["3.0", "z"]]


def _xlsx(path, sheet, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_compare_warning_blocks_match():
    out = os.path.join(tempfile.gettempdir(), "_skipwarn_cmp.xlsx")
    # No warnings → a clean match (sanity baseline).
    res = run_compare(SC, ROWS, ROWS, False, out, mode="values")
    assert res.verdict == "match", res.verdict
    os.remove(out)

    # With a skipped input the SAME identical data must NOT report a clean match.
    res = run_compare(SC, ROWS, ROWS, False, out, mode="values",
                      warnings=["BENV highway_sequence_route_099.xlsx: could "
                                "not open (BadZipFile)"])
    assert res.verdict == "diff", ("warning must block a clean match", res.verdict)
    assert "COULD NOT COMPARE" in res.summary_lines[0], res.summary_lines[0]
    assert any("route_099" in ln for ln in res.summary_lines), \
        ("skipped file must be listed in the summary", res.summary_lines)
    wb = load_workbook(out, data_only=False)
    banner = wb["Summary"]["B3"].value
    wb.close()
    os.remove(out)
    assert isinstance(banner, str) and banner.startswith("=IF("), \
        ("workbook banner must be guarded by the freshness formula", banner)
    assert "REGENERATE REQUIRED" in banner, banner
    assert "COULD NOT COMPARE EVERYTHING" in banner, banner
    assert "✓ EVERYTHING MATCHES" not in banner, banner


def test_consolidate_partial_and_allfail():
    d = Path(tempfile.mkdtemp())
    _xlsx(d / "highway_log_route_001.xlsx", "S", ["Loc", "V"], [["1", "a"]])
    _xlsx(d / "highway_log_route_002.xlsx", "WRONG", ["Loc", "V"], [["2", "b"]])

    # Partial: one good + one missing-sheet → status ok, but LOUD incomplete.
    out = d / "consolidated.xlsx"
    res = consolidate_xlsx(input_dir=d, out_path=out, sheet_name="S",
                           report_name="Highway Log", title="t")
    assert res.status == "ok", res.status
    assert res.summary_lines[0].startswith("⚠ INCOMPLETE"), res.summary_lines[0]
    assert out.exists(), "the good route should still have been written"

    # All-fail: only the bad file, and a pre-existing output present → error,
    # and the existing output is NOT overwritten.
    d2 = Path(tempfile.mkdtemp())
    _xlsx(d2 / "highway_log_route_002.xlsx", "WRONG", ["Loc", "V"], [["2", "b"]])
    out2 = d2 / "consolidated.xlsx"
    out2.write_bytes(b"PRIOR-GOOD-FILE")
    res = consolidate_xlsx(input_dir=d2, out_path=out2, sheet_name="S",
                           report_name="Highway Log", title="t")
    assert res.status == "error", ("all-fail must be an error", res.status)
    assert out2.read_bytes() == b"PRIOR-GOOD-FILE", \
        "all-fail must NOT overwrite the existing good file"


def main():
    test_compare_warning_blocks_match()
    test_consolidate_partial_and_allfail()
    print("OK  COMPARE-SKIPPED-FILES-MATCH + CONSOLIDATE-XLSX-PARTIAL-OK: "
          "warnings block a clean match (verdict=diff, banner not-a-match, "
          "files listed); consolidate flags partials loudly and won't "
          "overwrite on all-fail.")


if __name__ == "__main__":
    main()
