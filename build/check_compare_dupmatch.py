"""Golden check for COMPARE-DUPLICATE-KEY SIMILARITY PAIRING (compare_core.py).

The TSN-vs-TSMIS (and cross-environment) comparison used to pair rows that share
a key by FILE ORDER: first-with-first, second-with-second. When a key legitimately
repeats — the field report: two "001 R000.129" segments — that paired a row with
the WRONG twin and reported phantom differences, even though the same rows match
perfectly when paired the right way. compare_core now pairs duplicate-key rows by
data SIMILARITY (the most-alike rows share an occurrence #); the optimal pairing's
total diff count is <= any positional pairing's, so it can only REMOVE phantom
diffs, never add one.

Run with the build venv:
    build\\.venv\\Scripts\\python.exe build\\check_compare_dupmatch.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

from compare_core import (CompareSchema, _DIFF_MARK, count_diffs, keys_for,
                          pair_occurrences_by_similarity, run_compare, union_keys)
from openpyxl import load_workbook

_fail = []


def check(name, cond):
    print(f"  [{'OK ' if cond else 'FAIL'}] {name}")
    if not cond:
        _fail.append(name)


# Per-route Highway Log shape (no Route column): key on Location. Two rows share
# the key "R000.129"; side B lists them in the OPPOSITE order, so the truly-equal
# rows are NOT in the same file position.
HDR = ["Location", "Lanes", "Width"]
A = [["R000.129", "2", "12"], ["R000.129", "4", "24"]]
B = [["R000.129", "4", "24"], ["R000.129", "2", "12"]]   # same data, swapped order
SC = CompareSchema(report_name="Dup", header=HDR, side_a="TSMIS", side_b="TSN",
                   id_noun="location", id_noun_plural="locations",
                   sides_noun="systems")


def test_positional_was_wrong_similarity_fixes():
    print("duplicate key: file-order pairing is wrong, similarity pairing is right:")
    kt, kn = keys_for(A, False), keys_for(B, False)
    # 1) The bug: pairing by file order (occ 1<->1, 2<->2) reports 4 phantom diffs.
    c_pos = count_diffs(SC, A, B, kt, kn, union_keys(kt, kn), False)
    check("file-order pairing reports the phantom diffs (4 cells / 2 rows)",
          c_pos["both"] == 2 and c_pos["diff_cells"] == 4 and c_pos["diff_rows"] == 2)

    # 2) The fix: similarity pairing crosses the occurrences so the equal rows pair.
    mt, mn = pair_occurrences_by_similarity(SC, A, B, kt, kn, False)
    c_sim = count_diffs(SC, A, B, mt, mn, union_keys(mt, mn), False)
    check("similarity pairing eliminates the phantom diffs (0 cells / 0 rows)",
          c_sim["both"] == 2 and c_sim["diff_cells"] == 0 and c_sim["diff_rows"] == 0)
    check("still two matched rows, none one-sided",
          c_sim["t_only"] == 0 and c_sim["n_only"] == 0)
    # The occurrences crossed: A's first row (Lanes 2) now shares an occ with B's
    # SECOND row (Lanes 2), not B's first.
    check("matched occurrences crossed to the look-alike rows",
          mt[0][2] == mn[1][2] and mt[1][2] == mn[0][2] and mt[0][2] != mt[1][2])


def test_end_to_end_values_and_formulas():
    print("end-to-end run_compare: zero differing cells, verdict matches:")
    # Both flavors share the same Python mirror, so verdict == "match" (zero diff
    # cells AND zero one-sided) proves the count for BOTH. Only the values flavor
    # holds literal computed cells we can inspect without Excel.
    for mode in ("values", "formulas"):
        out = os.path.join(tempfile.gettempdir(), f"_dupmatch_{mode}.xlsx")
        res = run_compare(SC, A, B, False, out, mode=mode)
        check(f"{mode}: status ok", res.status == "ok")
        check(f"{mode}: verdict is match (no real differences)",
              res.verdict == "match")
        if mode == "values":
            wb = load_workbook(out, data_only=True)
            body = list(wb["Comparison"].iter_rows(values_only=True))[1:]
            wb.close()
            # per-route layout: A=Location B=# C=TSMIS Row D=TSN Row E=Status F=Diffs G..
            check("values: two Both rows in the Comparison",
                  len(body) == 2 and all(r[4] == "Both" for r in body))
            neq = sum(1 for r in body for v in r
                      if isinstance(v, str) and _DIFF_MARK in v)
            check("values: zero differing cells written", neq == 0)
        os.remove(out)


def test_uneven_group_best_match_wins():
    print("uneven duplicate group (2 vs 1): the lone row pairs with its look-alike:")
    # Two A rows at the same key, ONE B row that equals the SECOND A row. The B row
    # must pair with A's second row (not the first), leaving A's first one-sided.
    a = [["R1", "9", "9"], ["R1", "2", "2"]]
    b = [["R1", "2", "2"]]
    kt, kn = keys_for(a, False), keys_for(b, False)
    mt, mn = pair_occurrences_by_similarity(SC, a, b, kt, kn, False)
    c = count_diffs(SC, a, b, mt, mn, union_keys(mt, mn), False)
    check("one matched (identical) row, one A-only, zero diffs",
          c["both"] == 1 and c["t_only"] == 1 and c["n_only"] == 0
          and c["diff_cells"] == 0)
    check("the B row shares an occurrence with A's second (look-alike) row",
          mn[0][2] == mt[1][2] and mt[0][2] != mt[1][2])


def test_no_duplicates_unchanged():
    print("no duplicates: occurrences untouched (behavior identical to before):")
    a = [["R1", "1", "1"], ["R2", "2", "2"]]
    b = [["R1", "1", "9"], ["R2", "2", "2"]]
    kt, kn = keys_for(a, False), keys_for(b, False)
    mt, mn = pair_occurrences_by_similarity(SC, a, b, kt, kn, False)
    check("keys unchanged when every key is unique",
          list(mt) == kt and list(mn) == kn)
    c = count_diffs(SC, a, b, mt, mn, union_keys(mt, mn), False)
    check("the real single diff is still counted", c["diff_cells"] == 1)


def main():
    test_positional_was_wrong_similarity_fixes()
    test_end_to_end_values_and_formulas()
    test_uneven_group_best_match_wins()
    test_no_duplicates_unchanged()
    print()
    if _fail:
        print(f"FAILED: {len(_fail)} check(s): {_fail}")
        return 1
    print("ALL COMPARE-DUPLICATE-MATCH CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
