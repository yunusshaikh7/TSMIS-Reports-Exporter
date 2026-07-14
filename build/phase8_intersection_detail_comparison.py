#!/usr/bin/env python3
"""Independent Stage-8 Intersection Detail source and comparison oracle.

The truth path imports only audit-owned XLSX readers/normalizers and third-party
PDF primitives.  It never imports an application parser, comparator, schema, or
workbook result.  TSMIS PDF rows are reconstructed from word geometry against an
independently validated per-document grid; production reconstructs character
geometry against document-wide median windows, so the implementations do not
share extraction or pairing logic.

This file is intentionally built in two layers.  The source-oracle layer can run
alone while the expensive product witness is being developed.  The final Stage-8
acceptance layer will additionally invoke ``phase8_intersection_detail_product_witness``
and independently inspect all five formula/value workbook pairs.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
import hashlib
import json
import logging
import math
from pathlib import Path
import re
import subprocess
import sys
from typing import Iterable, Sequence
import zipfile
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl import load_workbook
import pdfplumber


BUILD_ROOT = Path(__file__).resolve().parent
REPO_ROOT = BUILD_ROOT.parent
GENERATOR_PATH = Path(__file__).resolve()
PRODUCT_HELPER_PATH = GENERATOR_PATH.with_name(
    "phase8_intersection_detail_product_witness.py")
SELF_GATE_PATH = GENERATOR_PATH.with_name(
    "check_phase8_intersection_detail_comparison.py")
sys.path.insert(0, str(BUILD_ROOT))

from phase3_independent_oracle import (  # noqa: E402
    FieldRule,
    OracleOutcome,
    OracleRow,
    OracleSchema,
    ValueRule,
    compare_rows,
    normalize_value,
)
from phase3_intersection_detail_oracle import (  # noqa: E402
    ASSERTED_FIELDS,
    SHARED_HEADER,
    TSMIS_HEADERS,
    TSMIS_SPEC,
    TSN_HEADERS,
    TSN_SPEC,
    _TSMIS_LOCATION_POSITION,
    _TSMIS_POSITION,
    _TSN_COLUMN,
    _TSN_INDEX,
    _project,
    normalize_pm,
)
from phase3_xlsx_stream import (  # noqa: E402
    SCALAR,
    ColumnSpec,
    SheetSpec,
    read_sheet,
)


SOURCE_ROOT = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\ground-truth\All Reports 7.9"
    r"\2026-07-09 ars-prod"
)
DEFAULT_TSMIS_XLSX_ROOT = SOURCE_ROOT / "intersection_detail"
DEFAULT_TSMIS_PDF_ROOT = SOURCE_ROOT / "intersection_detail_pdf"
DEFAULT_TSN_RAW = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\intersection_detail\raw"
    r"\TSAR - INTERSECTION DETAIL_TSN.xlsx"
)
DEFAULT_TSN_PDF = Path(
    r"C:\Users\Yunus\Downloads\TSMIS\tsn_library\intersection_detail\pdf"
    r"\Intersection Detail Statewide_TSN.pdf"
)
PHASE4_ROOT = Path(
    r"C:\Users\Yunus\.codex\visualizations\2026\07\10"
    r"\019f4e12-9bbf-7000-949a-019b80f60bdd\phase4_tsn_rebaseline"
)
R7_ROOT = (
    PHASE4_ROOT / "raw-2026-07-12-r7" / "intersection_detail" /
    "consolidated"
)
DEFAULT_TSN_NORMALIZED = R7_ROOT / "tsn_intersection_detail_normalized.xlsx"
DEFAULT_TSN_NORMALIZED_SIDECAR = Path(
    str(DEFAULT_TSN_NORMALIZED) + ".outcome.json")
DEFAULT_STAGE6_RESULT = (
    PHASE4_ROOT / "phase6_intersection_detail_conservation_r7_accepted.json")
DEFAULT_STAGE6_ACCEPTANCE = Path(
    str(DEFAULT_STAGE6_RESULT) + ".acceptance.json")
DEFAULT_TSN_CROSS_FORMAT = (
    PHASE4_ROOT / "intersection-tsn-cross-format-oracle-v2.json")
DEFAULT_PRODUCT_ROOT = REPO_ROOT / "tmp" / "phase8-intersection-detail-product-full-r1"

TREE_BINDINGS = {
    "tsmis_excel": {
        "suffix": ".xlsx", "files": 217, "bytes": 23_464_055,
        "manifest_sha256": (
            "885149005ab9a261ca83b686f68cfc3fc4fe550d8fd42d99252dcd36fb365bc9"),
    },
    "tsmis_pdf": {
        "suffix": ".pdf", "files": 217, "bytes": 31_673_183,
        "manifest_sha256": (
            "01e62eb195ab0bd5494cdb1b7a6a5ccbc35bd451bb5320a9bab0a045c58773c9"),
    },
}
FILE_BINDINGS = {
    "tsn_raw": {
        "bytes": 2_920_705,
        "sha256": "5170ab19b957ba78ab0f175571f3aab51e8c49cac13fa307b3d0beaa023c84a2",
    },
    "tsn_pdf": {
        "bytes": 9_284_543,
        "sha256": "1230b955176a1a34223ce8f79eeeed1b46970031372acc510ffb78a45c2f1f46",
    },
    "tsn_normalized": {
        "bytes": 2_084_691,
        "sha256": "d4609c3afb8663dd89e6e2e00103d41245a0213d7e4e08fb63e961bc4035b37b",
    },
    "tsn_normalized_sidecar": {
        "bytes": 903,
        "sha256": "9a62c3341d9c78dbab7c9eef01c23c714081499dd44cdeac85ef21b1f1c2a5b8",
    },
    "stage6_result": {
        "bytes": 453_532,
        "sha256": "4d507661835cdd9e9267f05f7700777ba97b8a3948797ac3e436be8db8d21b88",
    },
    "stage6_acceptance": {
        "bytes": 3_353,
        "sha256": "7077358da9ca016c12a4d1bc2cf8e09c95b20ac588272febf9b307f5856c7b43",
    },
    "tsn_cross_format": {
        "bytes": 91_032,
        "sha256": "63f5741203b06ef37245f195953058cf45ec921c04aaa00ccf676e44baba2c2e",
    },
}

NORMALIZED_HEADERS = (
    "Route", "PR", "Route Suffix", "PM", "Date of Record", "HG",
    "City Code", "R/U", "INT Type Eff-Date", "INT Type",
    "Control Type Eff-Date", "Control Type", "Lighting Eff-Date",
    "Lighting", "ML Eff-Date", "ML Mastarm", "ML Left Chan",
    "ML Right Chan", "ML Traffic Flow", "ML Num Lanes", "Description",
    "Main Line Length", "CS Eff-Date", "CS Mastarm", "CS Left Chan",
    "CS Right Chan", "CS Traffic Flow", "CS Num Lanes",
    "Int St Eff-Date", "Intrte Route", "Intrte PM Prefix",
    "Intrte Postmile", "Intrte PM Suffix", "Xing Line Lgth",
    "TSN District", "TSN County",
)
NORMALIZED_SPEC = SheetSpec(
    "Intersection Detail (TSN)",
    tuple(ColumnSpec(header, SCALAR) for header in NORMALIZED_HEADERS),
    exact_schema=True,
)
CONSOLIDATED_SPEC = SheetSpec(
    "Intersection Detail",
    (ColumnSpec("Route", SCALAR),
     *(ColumnSpec(header, SCALAR) for header in TSMIS_HEADERS)),
    exact_schema=True,
)

SOURCE_ASSERTED_FIELDS = ("District", "County", *ASSERTED_FIELDS)
PHYSICAL_SCHEMA = OracleSchema(
    key_rules=(
        ValueRule("Route"), ValueRule("County"),
        ValueRule("Complete PP"), ValueRule("Numeric PM"),
    ),
    field_rules=tuple(FieldRule(field, asserting=True)
                      for field in SOURCE_ASSERTED_FIELDS),
)
PRODUCT_SCHEMA = OracleSchema(
    key_rules=(ValueRule("Route"), ValueRule("PM")),
    field_rules=tuple(FieldRule(field, asserting=True)
                      for field in ASSERTED_FIELDS),
)

_EXCEL_TSN_PER_FIELD = {
    "CS Eff-Date": 509, "CS Left Chan": 153, "CS Mastarm": 155,
    "CS Num Lanes": 151, "CS Right Chan": 181, "CS Traffic Flow": 155,
    "City Code": 110, "Control Type": 1, "Control Type Eff-Date": 4,
    "Date of Record": 4, "Description": 4, "HG": 818,
    "INT Type": 1, "INT Type Eff-Date": 4, "Int St Eff-Date": 16_041,
    "Lighting Eff-Date": 4, "ML Eff-Date": 1_969,
    "ML Left Chan": 204, "ML Mastarm": 199, "ML Num Lanes": 186,
    "ML Right Chan": 190, "ML Traffic Flow": 189,
    "Main Line Length": 230, "R/U": 38, "Xing Line Lgth": 176,
}
_PDF_TSN_PER_FIELD = {
    **_EXCEL_TSN_PER_FIELD, "Description": 12, "HG": 817,
}
EXPECTED_PRODUCT_COUNTS = {
    "excel_vs_tsn_raw": {
        "paired_rows": 16_199, "side_a_only_rows": 260,
        "side_b_only_rows": 427, "differing_rows": 16_053,
        "differing_cells": 21_676, "asserted_cells": 518_368,
        "context_cells": 0, "per_field_counts": _EXCEL_TSN_PER_FIELD,
    },
    "excel_vs_tsn_normalized": {
        "paired_rows": 16_199, "side_a_only_rows": 260,
        "side_b_only_rows": 427, "differing_rows": 16_053,
        "differing_cells": 21_676, "asserted_cells": 518_368,
        "context_cells": 0, "per_field_counts": _EXCEL_TSN_PER_FIELD,
    },
    "pdf_vs_tsn_raw": {
        "paired_rows": 16_199, "side_a_only_rows": 260,
        "side_b_only_rows": 427, "differing_rows": 16_053,
        "differing_cells": 21_683, "asserted_cells": 518_368,
        "context_cells": 0, "per_field_counts": _PDF_TSN_PER_FIELD,
    },
    "pdf_vs_tsn_normalized": {
        "paired_rows": 16_199, "side_a_only_rows": 260,
        "side_b_only_rows": 427, "differing_rows": 16_053,
        "differing_cells": 21_683, "asserted_cells": 518_368,
        "context_cells": 0, "per_field_counts": _PDF_TSN_PER_FIELD,
    },
    "pdf_vs_excel": {
        "paired_rows": 16_459, "side_a_only_rows": 0,
        "side_b_only_rows": 0, "differing_rows": 9,
        "differing_cells": 9, "asserted_cells": 526_688,
        "context_cells": 0,
        "per_field_counts": {"Description": 8, "HG": 1},
    },
}
EXPECTED_SOURCE_COUNTS = {
    "excel_vs_tsn_raw": {
        "known": True, "paired_rows": 16_199,
        "side_a_only_rows": 260, "side_b_only_rows": 427,
        "differing_rows": 16_053, "differing_cells": 21_676,
        "per_field_counts": _EXCEL_TSN_PER_FIELD,
        "asserted_cells": 550_766, "context_cells": 0,
    },
    "pdf_vs_tsn_raw": {
        "known": True, "paired_rows": 16_199,
        "side_a_only_rows": 260, "side_b_only_rows": 427,
        "differing_rows": 16_053, "differing_cells": 21_683,
        "per_field_counts": _PDF_TSN_PER_FIELD,
        "asserted_cells": 550_766, "context_cells": 0,
    },
    "pdf_vs_excel": {
        "known": True, "paired_rows": 16_459,
        "side_a_only_rows": 0, "side_b_only_rows": 0,
        "differing_rows": 9, "differing_cells": 9,
        "per_field_counts": {"Description": 8, "HG": 1},
        "asserted_cells": 559_606, "context_cells": 0,
    },
    "raw_vs_normalized": {
        "known": True, "paired_rows": 16_626,
        "side_a_only_rows": 0, "side_b_only_rows": 0,
        "differing_rows": 0, "differing_cells": 0,
        "per_field_counts": {}, "asserted_cells": 565_284,
        "context_cells": 0,
    },
}

LOCATION_RE = re.compile(
    r"^\s*(\d{1,2})\s+([A-Z]{2,3})\.?\s+(\d+)([A-Z]?)\s*$",
    re.IGNORECASE,
)
MEMBER_RE = re.compile(
    r"^intersection_detail_route_(\d{3}[A-Z]?)\.(?:xlsx|pdf)$",
    re.IGNORECASE,
)
ROW_A_PM_RE = re.compile(r"^\d{3}\.\d{3}$")
OLD_PM_RE = re.compile(r"^\d{1,2}\.\d{3}$")
ROW_B_NUMBER_RE = re.compile(r"^\d+$")

# Current July-2026 reports dynamically size columns to the route's content.  A
# document must nevertheless have one exact 21-cell row-A profile and one exact
# 18-cell row-B profile; row B is row A with grid columns 3..6 merged.  We derive
# that one profile from raw rectangle bands, validate every band against it, and
# use it only to assign pdfplumber WORDS.  Production instead takes per-column
# medians and assigns CHARACTERS, so malformed tokenization/pairing remains an
# independently observable failure.
PDF_GRID_QUANTUM = 0.75
PDF_LEFT_EDGE = 27.75
PDF_RIGHT_EDGE = 764.25


class AuditError(RuntimeError):
    pass


@dataclass(frozen=True)
class FileEntry:
    name: str
    bytes: int
    sha256: str


@dataclass(frozen=True)
class DetailRow:
    source_index: int
    source: str
    source_ref: str
    route: str
    district: str
    county: str
    complete_pp: str
    pm: str
    numeric_pm: str
    route_suffix: str
    values: tuple[object, ...]
    member_route: str = ""
    physical_s: str = ""
    location_literal: str = ""
    source_only: tuple[tuple[str, object], ...] = ()
    raw_values: tuple[object, ...] = ()

    @property
    def physical_key(self) -> tuple[str, str, str, str]:
        return (self.route, self.county, self.complete_pp, self.numeric_pm)

    def oracle_row(self) -> OracleRow:
        return OracleRow(
            source_index=self.source_index,
            key=self.physical_key,
            values=(self.district, self.county, *self.values),
            source_ref=self.source_ref,
        )

    def product_oracle_row(self) -> OracleRow:
        return OracleRow(
            source_index=self.source_index,
            key=(self.route, self.pm),
            values=self.values,
            source_ref=self.source_ref,
        )

    def product_projection(self) -> tuple[object, ...]:
        asserted = dict(zip(ASSERTED_FIELDS, self.values))
        shared = tuple(
            self.pm if field == "PM" else asserted[field]
            for field in SHARED_HEADER)
        return (self.route, *shared)


def _sha_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _file_identity(path: Path) -> dict[str, object]:
    return {
        "path": str(path.resolve()), "bytes": path.stat().st_size,
        "sha256": _sha_file(path),
    }


def _canonical(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"),
        default=str,
    ).encode("utf-8")


def _manifest(root: Path, suffix: str) -> tuple[dict[str, object], list[FileEntry]]:
    paths = sorted(root.glob(f"*{suffix}"), key=lambda path: path.name)
    entries = [FileEntry(path.name, path.stat().st_size, _sha_file(path))
               for path in paths]
    wire = "".join(
        f"{entry.name}\t{entry.bytes}\t{entry.sha256}\n" for entry in entries
    ).encode("utf-8")
    return ({
        "files": len(entries),
        "bytes": sum(entry.bytes for entry in entries),
        "manifest_sha256": _sha_bytes(wire),
        "serialization": "name\\tbytes\\tsha256\\n sorted by name",
    }, entries)


def _bind_tree(label: str, root: Path) -> dict[str, object]:
    expected = TREE_BINDINGS[label]
    observed, entries = _manifest(root, str(expected["suffix"]))
    for field in ("files", "bytes", "manifest_sha256"):
        if observed[field] != expected[field]:
            raise AuditError(
                f"{label} {field} drift: {observed[field]!r} != {expected[field]!r}")
    return {
        "root": str(root.resolve()), "binding": dict(expected),
        "observed": observed, "members": [asdict(entry) for entry in entries],
    }


def _bind_file(label: str, path: Path) -> dict[str, object]:
    expected = FILE_BINDINGS[label]
    observed = {"bytes": path.stat().st_size, "sha256": _sha_file(path)}
    if observed != expected:
        raise AuditError(f"{label} identity drift: {observed!r} != {expected!r}")
    return {"path": str(path.resolve()), "binding": dict(expected),
            "observed": observed}


def _read_json(path: Path) -> dict[str, object]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise AuditError(f"cannot read bound JSON {path}: {exc}") from exc
    if not isinstance(payload, dict):
        raise AuditError(f"bound JSON is not an object: {path}")
    return payload


def _accepted_dependencies(args: argparse.Namespace) -> dict[str, object]:
    identities = {
        "tsn_normalized_sidecar": _bind_file(
            "tsn_normalized_sidecar", args.tsn_normalized_sidecar),
        "stage6_result": _bind_file("stage6_result", args.stage6_result),
        "stage6_acceptance": _bind_file(
            "stage6_acceptance", args.stage6_acceptance),
        "tsn_cross_format": _bind_file(
            "tsn_cross_format", args.tsn_cross_format),
    }
    sidecar = _read_json(args.tsn_normalized_sidecar)
    sidecar_checks = {
        "schema_version": sidecar.get("schema_version") == 1,
        "complete": (
            sidecar.get("completion"), sidecar.get("skipped_inputs"),
            sidecar.get("failed_inputs")) == ("complete", 0, 0),
        "normalization_version": sidecar.get("tsn_normalization_version") == 3,
        "raw_member_bound": (
            sidecar.get("tsn_raw_manifest", {}).get("members") == [{
                "relative_path": "TSAR - INTERSECTION DETAIL_TSN.xlsx",
                "byte_length": FILE_BINDINGS["tsn_raw"]["bytes"],
                "sha256": FILE_BINDINGS["tsn_raw"]["sha256"],
            }]),
        "normalized_member_bound": (
            sidecar.get("tsn_normalized_workbook_identity") == {
                "version": 1, "algorithm": "sha256",
                "byte_length": FILE_BINDINGS["tsn_normalized"]["bytes"],
                "sha256": FILE_BINDINGS["tsn_normalized"]["sha256"],
            }),
    }
    if not all(sidecar_checks.values()):
        raise AuditError(f"normalized outcome dependency drift: {sidecar_checks!r}")

    stage6 = _read_json(args.stage6_result)
    stage6_acceptance = _read_json(args.stage6_acceptance)
    blocking = stage6.get("findings", {}).get("blocking", [])
    blocking_fields = sorted(
        str(item.get("field")) for item in blocking if isinstance(item, dict))
    stage6_checks = {
        "acceptance_result_hash": (
            stage6_acceptance.get("result_bytes"),
            stage6_acceptance.get("result_sha256")) == (
                FILE_BINDINGS["stage6_result"]["bytes"],
                FILE_BINDINGS["stage6_result"]["sha256"]),
        "postwrite_revalidated": (
            stage6_acceptance.get("post_result_write_revalidation") is True),
        "all_24_invariants_true": (
            len(stage6.get("audit_invariants", {})) == 24
            and all(stage6.get("audit_invariants", {}).values())),
        "projection_exact": stage6.get("projection_exact") is True,
        "family_complete": stage6.get("stage6_family_audit_complete") is True,
        "known_three_source_claim_omissions_only": (
            stage6.get("normalized_full_conservation") is False
            and blocking_fields == ["CROSS_ADT", "MAIN_ADT", "MAIN_EFF_DATE"]),
    }
    if not all(stage6_checks.values()):
        raise AuditError(f"Stage-6 dependency drift: {stage6_checks!r}")

    cross = _read_json(args.tsn_cross_format)
    detail = cross.get("detail_cross_format", {})
    cross_checks = {
        "status_pass": cross.get("status") == "pass",
        "unresolved_zero": cross.get("unresolved_gate_count") == 0,
        "source_contract_pass": cross.get("source_contract", {}).get("pass") is True,
        "negative_self_check_pass": (
            cross.get("internal_negative_mutation_self_check", {}).get("pass") is True),
        "all_16626_records_paired": (
            detail.get("xlsx_records"), detail.get("pdf_records"),
            detail.get("paired_records")) == (16_626, 16_626, 16_626),
        "all_598536_cells_classified": detail.get("asserted_cells") == 598_536,
        "one_bound_source_export_delta": (
            detail.get("relation_counts") == {
                "exact": 578_432, "render_equivalent": 20_103,
                "source_export_delta": 1}),
        "physical_identity_exact": (
            detail.get("physical_identity", {}).get("definition") == [
                "base_route", "county", "complete_PP", "numeric_POST_MILE"]
            and detail.get("physical_identity", {}).get(
                "xlsx_unique_identities") == 16_611
            and detail.get("physical_identity", {}).get(
                "pdf_unique_identities") == 16_611),
        "report_view_mapping_complete": (
            cross.get("report_view_source_mapping", {}).get(
                "all_xlsx_columns_mapped_once") is True),
    }
    if not all(cross_checks.values()):
        raise AuditError(f"TSN cross-format dependency drift: {cross_checks!r}")
    return {
        "identities": identities,
        "normalized_outcome": {"checks": sidecar_checks, "payload": sidecar},
        "stage6_raw_to_normalized": {
            "checks": stage6_checks,
            "blocking_fields": blocking_fields,
            "projection_comparison": stage6.get("projection_comparison"),
            "identity_and_collision_census": stage6.get(
                "identity_and_collision_census"),
        },
        "tsn_xlsx_to_pdf": {
            "checks": cross_checks,
            "detail_cross_format": detail,
            "report_view_source_mapping": cross.get(
                "report_view_source_mapping"),
        },
    }


def _text(value: object) -> str:
    if value is None:
        return ""
    if type(value) is bool:
        return "TRUE" if value else "FALSE"
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise AuditError("non-finite Decimal source value")
        return format(value, "f")
    return str(value).strip()


def _location(value: object) -> tuple[str, str, str, str]:
    literal = _text(value).upper().replace("-", " ")
    match = LOCATION_RE.fullmatch(literal)
    if match is None:
        raise AuditError(f"invalid Intersection Detail Location: {literal!r}")
    district, county, route, suffix = match.groups()
    return f"{int(district):02d}", county.rstrip("."), f"{int(route):03d}", suffix


def _numeric_pm(value: object) -> str:
    literal = normalize_pm(value)
    try:
        number = Decimal(literal)
    except InvalidOperation as exc:
        raise AuditError(f"invalid numeric postmile: {literal!r}") from exc
    if not number.is_finite():
        raise AuditError(f"non-finite numeric postmile: {literal!r}")
    if number == 0:
        return "0"
    rendered = format(number.normalize(), "f")
    return rendered.rstrip("0").rstrip(".") if "." in rendered else rendered


def _detail_row_from_tsmis(
        values: Sequence[object], *, source_index: int, source: str,
        source_ref: str, member_route: str) -> DetailRow:
    if len(values) != len(TSMIS_HEADERS):
        raise AuditError(f"TSMIS row width {len(values)} != {len(TSMIS_HEADERS)}")
    location_literal = _text(values[_TSMIS_LOCATION_POSITION])
    district, county, route, suffix = _location(location_literal)
    projected = {
        field: (suffix if field == "Route Suffix"
                else _project(field, values[_TSMIS_POSITION[field]]))
        for field in SHARED_HEADER
    }
    return DetailRow(
        source_index=source_index, source=source, source_ref=source_ref,
        route=route, district=district, county=county,
        complete_pp=_text(projected["PR"]), pm=_text(projected["PM"]),
        numeric_pm=_numeric_pm(values[_TSMIS_POSITION["PM"]]),
        route_suffix=suffix,
        values=tuple(projected[field] for field in ASSERTED_FIELDS),
        member_route=member_route, physical_s=_text(values[2]),
        location_literal=location_literal,
        raw_values=tuple(values),
    )


def _detail_row_from_tsn(
        values: Sequence[object], *, source_index: int,
        source_ref: str) -> DetailRow:
    if len(values) != len(TSN_HEADERS):
        raise AuditError(f"TSN row width {len(values)} != {len(TSN_HEADERS)}")
    location_literal = _text(values[_TSN_INDEX["LOCATION"]])
    district, county, route, suffix = _location(location_literal)
    projected = {
        field: (suffix if field == "Route Suffix"
                else _project(field, values[_TSN_INDEX[_TSN_COLUMN[field]]]))
        for field in SHARED_HEADER
    }
    claims = tuple((field, values[_TSN_INDEX[field]]) for field in (
        "MAIN_EFF_DATE", "MAIN_ADT", "CROSS_ADT"))
    return DetailRow(
        source_index=source_index, source="TSN raw XLSX", source_ref=source_ref,
        route=route, district=district, county=county,
        complete_pp=_text(projected["PR"]), pm=_text(projected["PM"]),
        numeric_pm=_numeric_pm(values[_TSN_INDEX["POST_MILE"]]),
        route_suffix=suffix,
        values=tuple(projected[field] for field in ASSERTED_FIELDS),
        location_literal=location_literal,
        source_only=claims,
        raw_values=tuple(values),
    )


def _detail_row_from_normalized(
        values: Sequence[object], *, source_index: int,
        source_ref: str) -> DetailRow:
    if len(values) != len(NORMALIZED_HEADERS):
        raise AuditError(
            f"normalized row width {len(values)} != {len(NORMALIZED_HEADERS)}")
    raw = dict(zip(NORMALIZED_HEADERS, values))
    route = f"{int(_text(raw['Route'])):03d}"
    projected = {
        field: _project(field, raw[field]) for field in SHARED_HEADER
    }
    suffix = _text(projected["Route Suffix"]).upper()
    return DetailRow(
        source_index=source_index, source="TSN normalized r7",
        source_ref=source_ref, route=route,
        district=f"{int(_text(raw['TSN District'])):02d}",
        county=_text(raw["TSN County"]).upper().rstrip("."),
        complete_pp=_text(projected["PR"]), pm=_text(projected["PM"]),
        numeric_pm=_numeric_pm(raw["PM"]), route_suffix=suffix,
        values=tuple(projected[field] for field in ASSERTED_FIELDS),
        location_literal=(
            f"{int(_text(raw['TSN District'])):02d} "
            f"{_text(raw['TSN County']).upper()} {route}{suffix}"),
        raw_values=tuple(values),
    )


def _row_payload(row: DetailRow) -> tuple[object, ...]:
    return (
        row.route, row.district, row.county, row.complete_pp, row.pm,
        row.numeric_pm, row.route_suffix, row.member_route, row.physical_s,
        row.location_literal, *row.values, *row.source_only,
        *row.raw_values,
    )


def _rows_digest(rows: Sequence[DetailRow]) -> str:
    digest = hashlib.sha256()
    for row in rows:
        digest.update(_canonical(_row_payload(row)))
        digest.update(b"\n")
    return digest.hexdigest()


def _member_token(path: Path) -> str:
    match = MEMBER_RE.fullmatch(path.name)
    if match is None:
        raise AuditError(f"unexpected source member name: {path.name}")
    token = match.group(1).upper()
    base = f"{int(token[:3]):03d}"
    return base + token[3:]


def _parse_tsmis_excel(root: Path) -> dict[str, object]:
    rows: list[DetailRow] = []
    member_routes = []
    provenance_mismatches = []
    for path in sorted(root.glob("*.xlsx"), key=lambda item: item.name):
        member = _member_token(path)
        member_routes.append(member)
        sheet = read_sheet(path, TSMIS_SPEC)
        for physical in sheet.rows:
            if not any(value not in (None, "") for value in physical.values):
                raise AuditError(f"blank physical TSMIS row: {path.name}:{physical.source_row}")
            row = _detail_row_from_tsmis(
                physical.values, source_index=len(rows), source="TSMIS Excel",
                source_ref=f"{path.name}:row {physical.source_row}",
                member_route=member)
            derived = row.route + row.route_suffix
            if derived != member:
                provenance_mismatches.append({
                    "member": member, "derived": derived,
                    "source_ref": row.source_ref,
                })
            rows.append(row)
    return _source_summary(
        "TSMIS Excel", rows, member_routes,
        {"member_route_mismatches": provenance_mismatches})


def _parse_tsn_raw(path: Path) -> dict[str, object]:
    sheet = read_sheet(path, TSN_SPEC)
    rows = []
    for physical in sheet.rows:
        if not any(value not in (None, "") for value in physical.values):
            raise AuditError(f"blank TSN physical row {physical.source_row}")
        rows.append(_detail_row_from_tsn(
            physical.values, source_index=len(rows),
            source_ref=f"{path.name}:row {physical.source_row}"))
    return _source_summary("TSN raw XLSX", rows, sorted({row.route for row in rows}), {
        "source_only_claim_digests": {
            field: _sha_bytes(_canonical([
                dict(row.source_only)[field] for row in rows
            ])) for field in ("MAIN_EFF_DATE", "MAIN_ADT", "CROSS_ADT")
        },
        "source_only_nonblank_counts": {
            field: sum(dict(row.source_only)[field] not in (None, "")
                       for row in rows)
            for field in ("MAIN_EFF_DATE", "MAIN_ADT", "CROSS_ADT")
        },
    })


def _parse_tsn_normalized(path: Path) -> dict[str, object]:
    sheet = read_sheet(path, NORMALIZED_SPEC)
    rows = []
    for physical in sheet.rows:
        if not any(value not in (None, "") for value in physical.values):
            raise AuditError(f"blank normalized physical row {physical.source_row}")
        rows.append(_detail_row_from_normalized(
            physical.values, source_index=len(rows),
            source_ref=f"{path.name}:row {physical.source_row}"))
    return _source_summary(
        "TSN normalized r7", rows, sorted({row.route for row in rows}),
        {"source_only_columns_present": []})


def _rect_band_profiles(page) -> list[tuple[int, tuple[float, ...], float]]:
    by_top: defaultdict[float, list[dict[str, object]]] = defaultdict(list)
    for rect in page.rects:
        width = rect["x1"] - rect["x0"]
        height = rect["bottom"] - rect["top"]
        if 3 < width < page.width - 10 and 3 < height < 40:
            by_top[round(float(rect["top"]), 1)].append(rect)
    profiles = []
    for top, cells in sorted(by_top.items()):
        if len(cells) not in (21, 18):
            continue
        ordered = sorted(cells, key=lambda item: item["x0"])
        edges = tuple(round(float(value), 3) for value in (
            [ordered[0]["x0"]] + [item["x1"] for item in ordered]))
        profiles.append((len(cells), edges, top))
    return profiles


def _validate_pdf_profiles(
        path: Path, profiles: dict[int, Counter[tuple[float, ...]]]
        ) -> tuple[tuple[float, ...], tuple[float, ...]]:
    if set(profiles) != {21, 18}:
        raise AuditError(f"{path.name} PDF grid classes drifted: {sorted(profiles)}")
    if len(profiles[21]) != 1 or len(profiles[18]) != 1:
        raise AuditError(
            f"{path.name} has nonuniform PDF grids: "
            f"rowA={len(profiles[21])} rowB={len(profiles[18])}")
    a_edges = next(iter(profiles[21]))
    b_edges = next(iter(profiles[18]))
    if len(a_edges) != 22 or len(b_edges) != 19:
        raise AuditError(f"{path.name} PDF edge count drifted")
    expected_b = (*a_edges[:4], *a_edges[7:])
    if b_edges != expected_b:
        raise AuditError(
            f"{path.name} row-B grid is not the row-A 3..6 merge")
    for label, edges in (("rowA", a_edges), ("rowB", b_edges)):
        if edges[0] != PDF_LEFT_EDGE or edges[-1] != PDF_RIGHT_EDGE:
            raise AuditError(f"{path.name} {label} outer grid edges drifted")
        if any(right <= left for left, right in zip(edges, edges[1:])):
            raise AuditError(f"{path.name} {label} grid is not strictly increasing")
        if any(abs(edge / PDF_GRID_QUANTUM - round(edge / PDF_GRID_QUANTUM)) > 0.001
               for edge in edges):
            raise AuditError(f"{path.name} {label} grid left 0.75-point lattice")
    return a_edges, b_edges


def _word_lines(page, tolerance: float = 2.25) -> list[list[dict[str, object]]]:
    words = page.extract_words(
        x_tolerance=1, y_tolerance=2, keep_blank_chars=False,
        use_text_flow=False)
    ordered = sorted(words, key=lambda word: (float(word["top"]), float(word["x0"])))
    clusters: list[list[dict[str, object]]] = []
    anchors: list[float] = []
    for word in ordered:
        top = float(word["top"])
        if not clusters or abs(top - anchors[-1]) > tolerance:
            clusters.append([word])
            anchors.append(top)
        else:
            clusters[-1].append(word)
            anchors[-1] = sum(float(item["top"]) for item in clusters[-1]) / len(clusters[-1])
    return [sorted(cluster, key=lambda word: float(word["x0"]))
            for cluster in clusters]


def _assign_words(words: Sequence[dict[str, object]],
                  edges: Sequence[float]) -> list[str]:
    cells: list[list[dict[str, object]]] = [[] for _ in range(len(edges) - 1)]
    for word in words:
        center = (float(word["x0"]) + float(word["x1"])) / 2
        index = next((i for i in range(len(edges) - 1)
                      if edges[i] <= center < edges[i + 1]), None)
        if index is None:
            if center < edges[0] or center > edges[-1]:
                continue
            index = len(edges) - 2
        cells[index].append(word)
    return [
        " ".join(str(word["text"]) for word in sorted(
            members, key=lambda word: float(word["x0"]))).strip()
        for members in cells
    ]


def _pdf_row(a: Sequence[str], b: Sequence[str]) -> tuple[object, ...]:
    if len(a) != 21 or len(b) != 18:
        raise AuditError("PDF logical row widths drifted")
    if a[20]:
        raise AuditError(f"PDF vestigial row-A column became nonblank: {a[20]!r}")
    values = (
        *a[0:20], b[3], b[4], b[5], b[6], b[7], b[8], b[9], b[10],
        b[11], b[13], b[12], b[14], b[15], b[16], b[17],
    )
    return tuple(value or None for value in values)


def _parse_one_pdf(path: Path, start_index: int) -> tuple[
        list[DetailRow], dict[str, object]]:
    rows: list[DetailRow] = []
    pending: tuple[list[str], int] | None = None
    pages = row_a = row_b = old_layout = orphans = 0
    with pdfplumber.open(path) as document:
        pages = len(document.pages)
        profiles: dict[int, Counter[tuple[float, ...]]] = {
            21: Counter(), 18: Counter()}
        for page in document.pages:
            for width, edges, _top in _rect_band_profiles(page):
                profiles[width][edges] += 1
        a_edges, b_edges = _validate_pdf_profiles(path, profiles)
        a_bands = sum(profiles[21].values())
        b_bands = sum(profiles[18].values())
        for page_number, page in enumerate(document.pages, 1):
            for words in _word_lines(page):
                first = _assign_words(words, a_edges)
                if ROW_A_PM_RE.fullmatch(first[1]) and LOCATION_RE.search(first[3]):
                    row_a += 1
                    if pending is not None:
                        orphans += 1
                    pending = (first, page_number)
                    continue
                if OLD_PM_RE.fullmatch(first[1]) and LOCATION_RE.search(first[3]):
                    old_layout += 1
                    continue
                if pending is None:
                    continue
                second = _assign_words(words, b_edges)
                if ROW_B_NUMBER_RE.fullmatch(second[1]):
                    row_b += 1
                    values = _pdf_row(pending[0], second)
                    rows.append(_detail_row_from_tsmis(
                        values, source_index=start_index + len(rows),
                        source="TSMIS PDF",
                        source_ref=(f"{path.name}:pages {pending[1]}-{page_number}:"
                                    f"intersection {second[1]}"),
                        member_route=_member_token(path)))
                    pending = None
    if pending is not None:
        orphans += 1
    return rows, {
        "pages": pages, "row_a": row_a, "row_b": row_b,
        "orphans": orphans, "old_layout_rows": old_layout,
        "row_a_bands": a_bands, "row_b_bands": b_bands,
        "row_a_grid": list(a_edges), "row_b_grid": list(b_edges),
        "grid_profile_sha256": _sha_bytes(_canonical([a_edges, b_edges])),
    }


def _parse_tsmis_pdf(root: Path) -> dict[str, object]:
    logging.disable(logging.CRITICAL)
    rows: list[DetailRow] = []
    member_routes = []
    per_file = []
    grid_profiles = []
    provenance_mismatches = []
    totals = Counter()
    for path in sorted(root.glob("*.pdf"), key=lambda item: item.name):
        member = _member_token(path)
        member_routes.append(member)
        parsed, stats = _parse_one_pdf(path, len(rows))
        for row in parsed:
            derived = row.route + row.route_suffix
            if derived != member:
                provenance_mismatches.append({
                    "member": member, "derived": derived,
                    "source_ref": row.source_ref,
                })
        rows.extend(parsed)
        grid_profiles.append({
            "member": path.name,
            "row_a_grid": stats["row_a_grid"],
            "row_b_grid": stats["row_b_grid"],
            "sha256": stats["grid_profile_sha256"],
        })
        for key in ("pages", "row_a", "row_b", "orphans", "old_layout_rows",
                    "row_a_bands", "row_b_bands"):
            totals[key] += int(stats[key])
        if stats["orphans"] or stats["old_layout_rows"]:
            per_file.append({"member": path.name, **stats})
    if totals["orphans"] or totals["old_layout_rows"]:
        raise AuditError(f"PDF row reconciliation failed: {dict(totals)!r}")
    return _source_summary("TSMIS PDF", rows, member_routes, {
        "pdf_reconciliation": dict(totals),
        "problem_members": per_file,
        "grid_profiles": grid_profiles,
        "grid_profile_manifest_sha256": _sha_bytes(_canonical(grid_profiles)),
        "distinct_grid_profiles": len({item["sha256"] for item in grid_profiles}),
        "member_route_mismatches": provenance_mismatches,
        "extraction_independence": {
            "product": "characters assigned to document-wide median grid windows",
            "oracle": "words assigned to one exact, uniform per-document grid",
            "additional_oracle_contract": (
                "row-B grid must exactly merge row-A columns 3..6; every edge "
                "must remain on the 0.75-point lattice"),
        },
    })


def _identity_census(rows: Sequence[DetailRow]) -> dict[str, object]:
    strong = Counter(row.physical_key for row in rows)
    lossless = Counter((row.district, row.county, row.route,
                        row.route_suffix, row.complete_pp, row.numeric_pm)
                       for row in rows)
    route_pm_counties: defaultdict[tuple[str, str], set[str]] = defaultdict(set)
    route_pp_pm_counties: defaultdict[tuple[str, str, str], set[str]] = defaultdict(set)
    within: defaultdict[tuple[str, str, str], set[str]] = defaultdict(set)
    for row in rows:
        route_pm_counties[(row.route, row.numeric_pm)].add(row.county)
        route_pp_pm_counties[(row.route, row.complete_pp, row.numeric_pm)].add(row.county)
        within[(row.route, row.county, row.numeric_pm)].add(row.complete_pp)

    def multiplicity(counter: Counter[tuple[str, ...]]) -> dict[str, int]:
        duplicates = [count for count in counter.values() if count > 1]
        return {
            "unique": len(counter), "duplicate_groups": len(duplicates),
            "duplicate_occurrences": sum(duplicates),
            "max_multiplicity": max(counter.values(), default=0),
        }

    cross_route_pm = [counties for counties in route_pm_counties.values()
                      if len(counties) > 1]
    cross_route_pp_pm = [counties for counties in route_pp_pm_counties.values()
                         if len(counties) > 1]
    pp_collisions = [
        {"route": key[0], "county": key[1], "numeric_pm": key[2],
         "complete_pp_values": sorted(values)}
        for key, values in sorted(within.items()) if len(values) > 1
    ]
    return {
        "physical_identity": multiplicity(strong),
        "lossless_identity": multiplicity(lossless),
        "route_plus_numeric_pm_cross_county": {
            "keys": len(cross_route_pm),
            "county_identities": sum(len(value) for value in cross_route_pm),
        },
        "route_plus_complete_pp_plus_pm_cross_county": {
            "keys": len(cross_route_pp_pm),
            "county_identities": sum(len(value) for value in cross_route_pp_pm),
        },
        "within_county_route_pm_complete_pp_collisions": pp_collisions,
    }


def _source_summary(label: str, rows: Sequence[DetailRow],
                    member_routes: Sequence[str],
                    extra: dict[str, object]) -> dict[str, object]:
    return {
        "label": label, "rows": len(rows),
        "member_routes": list(member_routes),
        "member_route_count": len(member_routes),
        "derived_routes": sorted({row.route + row.route_suffix for row in rows}),
        "derived_route_count": len({row.route + row.route_suffix for row in rows}),
        "ordered_source_payload_sha256": _rows_digest(rows),
        "identity": _identity_census(rows),
        "district_census": dict(sorted(Counter(row.district for row in rows).items())),
        "county_census": dict(sorted(Counter(row.county for row in rows).items())),
        "route_suffix_census": dict(sorted(Counter(row.route_suffix for row in rows).items())),
        "member_route_census": dict(sorted(Counter(
            row.member_route for row in rows if row.member_route).items())),
        "physical_s_census": dict(sorted(Counter(
            row.physical_s for row in rows if row.member_route).items())),
        "explicit_tsmis_claims_sha256": _sha_bytes(_canonical([
            [row.member_route, row.physical_s, row.location_literal]
            for row in rows if row.member_route
        ])),
        "rows_data": list(rows),
        **extra,
    }


def _public_source(summary: dict[str, object]) -> dict[str, object]:
    return {key: value for key, value in summary.items() if key != "rows_data"}


def _comparison(label: str, left: Sequence[DetailRow],
                right: Sequence[DetailRow]) -> dict[str, object]:
    outcome = compare_rows(
        PHYSICAL_SCHEMA,
        tuple(row.oracle_row() for row in left),
        tuple(row.oracle_row() for row in right),
    )
    by_left = {row.source_index: row for row in left}
    by_right = {row.source_index: row for row in right}
    digest = hashlib.sha256()
    examples = []
    special = []
    for result in outcome.row_results:
        left_row = by_left[result.source_index_a]
        right_row = by_right[result.source_index_b]
        entry = {
            "key": list(left_row.physical_key),
            "left_ref": left_row.source_ref,
            "right_ref": right_row.source_ref,
            "differing_fields": list(result.differing_fields),
            "differences": [
                {
                    "field": field,
                    "left": cell.normalized_a.text if cell.normalized_a.kind == "text" else "",
                    "right": cell.normalized_b.text if cell.normalized_b.kind == "text" else "",
                }
                for field, cell in zip(SOURCE_ASSERTED_FIELDS, result.cells)
                if cell.counts_as_difference
            ],
        }
        digest.update(_canonical(entry))
        digest.update(b"\n")
        if result.differing_fields and len(examples) < 100:
            examples.append(entry)
        if left_row.route == "108" and left_row.numeric_pm == "5.87":
            special.append(entry)
    left_only = [
        {"key": list(by_left[index].physical_key),
         "source_ref": by_left[index].source_ref}
        for index in outcome.side_a_only_indices
    ]
    right_only = [
        {"key": list(by_right[index].physical_key),
         "source_ref": by_right[index].source_ref}
        for index in outcome.side_b_only_indices
    ]
    counts = asdict(outcome.counts)
    return {
        "label": label,
        "completion": outcome.completion,
        "verdict": outcome.verdict,
        "counts": counts,
        "pairing_quality": outcome.pairing_quality,
        "duplicate_pair_groups": sum(
            trace.side_a_size > 1 or trace.side_b_size > 1
            for trace in outcome.pairing_trace),
        "max_pair_matrix_cells": max(
            (trace.matrix_cells for trace in outcome.pairing_trace), default=0),
        "capped_diagnostics": [asdict(item) for item in outcome.capped_diagnostics],
        "ordered_pair_ledger_sha256": digest.hexdigest(),
        "difference_examples_first_100": examples,
        "route_108_pm_5_870_witness": special,
        "left_only": left_only,
        "right_only": right_only,
    }


def _product_text(value: object) -> str:
    normalized = normalize_value(value)
    return "" if normalized.kind == "blank" else normalized.text


def _counter_digest(counter: Counter[tuple[object, ...]]) -> str:
    ordered = sorted(counter.items(), key=lambda item: _canonical(item[0]))
    return _sha_bytes(_canonical([[key, count] for key, count in ordered]))


def _counter_difference_examples(
        observed: Counter[tuple[object, ...]],
        expected: Counter[tuple[object, ...]]) -> dict[str, object]:
    missing = expected - observed
    extra = observed - expected

    def first(counter: Counter[tuple[object, ...]]) -> list[dict[str, object]]:
        return [
            {"entry": list(key), "count": count}
            for key, count in sorted(
                counter.items(), key=lambda item: _canonical(item[0]))[:10]
        ]

    return {"missing_first_10": first(missing), "extra_first_10": first(extra)}


def _product_expected(
        label: str, side_a: str, side_b: str,
        left: Sequence[DetailRow], right: Sequence[DetailRow]) -> dict[str, object]:
    outcome = compare_rows(
        PRODUCT_SCHEMA,
        tuple(row.product_oracle_row() for row in left),
        tuple(row.product_oracle_row() for row in right),
    )
    if outcome.completion != "complete" or outcome.pairing_quality != "exact":
        raise AuditError(f"{label}: weak product oracle did not complete exactly")
    counts = asdict(outcome.counts)
    frozen = EXPECTED_PRODUCT_COUNTS[label]
    observed_frozen = {
        key: counts[key] for key in (
            "paired_rows", "side_a_only_rows", "side_b_only_rows",
            "differing_rows", "differing_cells", "asserted_cells",
            "context_cells", "per_field_counts")
    }
    if observed_frozen != frozen:
        raise AuditError(
            f"{label}: independently derived weak product counts drifted: "
            f"{observed_frozen!r}")

    paired: Counter[tuple[object, ...]] = Counter()
    for result in outcome.row_results:
        displays = []
        state = []
        for cell in result.cells:
            if cell.equal:
                displays.append(
                    "" if cell.normalized_a.kind == "blank"
                    else cell.normalized_a.text)
            else:
                displays.append(cell.display)
            state.append("E" if cell.equal else "D")
        key = tuple(
            "" if item.kind == "blank" else item.text for item in result.key)
        paired[(
            *key, len(result.differing_fields), *displays, "".join(state)
        )] += 1

    by_left = {row.source_index: row for row in left}
    by_right = {row.source_index: row for row in right}

    def one_sided(indices: Sequence[int], rows: dict[int, DetailRow]
                  ) -> Counter[tuple[object, ...]]:
        result: Counter[tuple[object, ...]] = Counter()
        for index in indices:
            row = rows[index]
            result[(
                _product_text(row.route), _product_text(row.pm),
                *(_product_text(value) for value in row.values),
            )] += 1
        return result

    side_a_only = one_sided(outcome.side_a_only_indices, by_left)
    side_b_only = one_sided(outcome.side_b_only_indices, by_right)
    return {
        "label": label, "side_a": side_a, "side_b": side_b,
        "counts": counts,
        "pairing": {
            "quality": outcome.pairing_quality,
            "groups": len(outcome.pairing_trace),
            "duplicate_groups": sum(
                trace.side_a_size > 1 or trace.side_b_size > 1
                for trace in outcome.pairing_trace),
            "max_matrix_cells": max(
                (trace.matrix_cells for trace in outcome.pairing_trace), default=0),
        },
        "paired_ledger_sha256": _counter_digest(paired),
        "side_a_only_ledger_sha256": _counter_digest(side_a_only),
        "side_b_only_ledger_sha256": _counter_digest(side_b_only),
        "_paired": paired,
        "_side_a_only": side_a_only,
        "_side_b_only": side_b_only,
    }


def _public_product_expected(expected: dict[str, object]) -> dict[str, object]:
    return {key: value for key, value in expected.items()
            if not key.startswith("_")}


def _inspect_consolidated(
        path: Path, truth_rows: Sequence[DetailRow]) -> dict[str, object]:
    sheet = read_sheet(path, CONSOLIDATED_SPEC)
    expected = [
        (row.member_route, *row.raw_values) for row in truth_rows
    ]
    if len(sheet.rows) != len(expected):
        raise AuditError(
            f"{path.name}: consolidated row count {len(sheet.rows)} != "
            f"{len(expected)}")
    observed_digest = hashlib.sha256()
    expected_digest = hashlib.sha256()
    first_mismatch = None
    member_route_mismatches = 0
    physical_s_mismatches = 0
    blank_serialization = Counter()

    def typed(values: Sequence[object]) -> list[tuple[str, str]]:
        return [(type(value).__name__, "" if value is None else str(value))
                for value in values]

    for index, (physical, wanted) in enumerate(zip(sheet.rows, expected), 2):
        actual = tuple(physical.values)
        observed_digest.update(_canonical(typed(actual)))
        observed_digest.update(b"\n")
        expected_digest.update(_canonical(typed(wanted)))
        expected_digest.update(b"\n")
        for column, (observed, required) in enumerate(zip(actual, wanted)):
            exact = type(observed) is type(required) and observed == required
            blank_equivalent = required == "" and observed is None
            if blank_equivalent:
                label = "Route" if column == 0 else TSMIS_HEADERS[column - 1]
                blank_serialization[label] += 1
            elif not exact and first_mismatch is None:
                first_mismatch = {
                    "row": index, "column": column + 1,
                    "header": "Route" if column == 0 else TSMIS_HEADERS[column - 1],
                    "observed_type": type(observed).__name__,
                    "observed": observed,
                    "expected_type": type(required).__name__,
                    "expected": required,
                }
        member_route_mismatches += int(
            _text(actual[0]) != truth_rows[index - 2].member_route)
        physical_s_mismatches += int(
            _text(actual[3]) != truth_rows[index - 2].physical_s)
    semantic_exact = first_mismatch is None
    if not semantic_exact:
        raise AuditError(
            f"{path.name}: product consolidation changed source bytes at "
            f"row {first_mismatch['row']} column {first_mismatch['column']}: "
            f"{first_mismatch!r}")
    return {
        "path": str(path.resolve()), "rows": len(expected),
        "columns": len(CONSOLIDATED_SPEC.columns),
        "ordered_source_payload_sha256": observed_digest.hexdigest(),
        "expected_ordered_source_payload_sha256": expected_digest.hexdigest(),
        "projection_exact": semantic_exact,
        "all_nonblank_typed_cells_exact": semantic_exact,
        "raw_representation_exact": not blank_serialization,
        "blank_string_to_physical_blank_cells": sum(blank_serialization.values()),
        "blank_string_to_physical_blank_by_column": dict(blank_serialization),
        "only_serialization_equivalence": "explicit empty string to physical blank",
        "explicit_member_route_mismatches": member_route_mismatches,
        "explicit_physical_s_mismatches": physical_s_mismatches,
    }


def _formula_tag_count(path: Path) -> int:
    return sum(_formula_tag_census(path).values())


def _formula_tag_census(path: Path) -> dict[str, int]:
    pattern = re.compile(rb"<(?:[A-Za-z0-9_]+:)?f(?:\s|>)")
    counts: dict[str, int] = {}
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(
            archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships}
        ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        for sheet in workbook.findall("main:sheets/main:sheet", ns):
            label = sheet.attrib["name"]
            relation = sheet.attrib[f"{{{ns['rel']}}}id"]
            target = targets[relation].replace("\\", "/")
            member_name = (target.lstrip("/") if target.startswith("/")
                           else "xl/" + target)
            count = 0
            tail = b""
            with archive.open(member_name) as member:
                for chunk in iter(lambda: member.read(1024 * 1024), b""):
                    payload = tail + chunk
                    count += sum(
                        match.end() > len(tail)
                        for match in pattern.finditer(payload))
                    tail = payload[-96:]
            counts[label] = count
    return counts


def _validate_file_identity(
        identity: dict[str, object], expected_path: Path, label: str
        ) -> dict[str, object]:
    actual = {
        "path": str(expected_path), "bytes": expected_path.stat().st_size,
        "sha256": _sha_file(expected_path),
    }
    declared_path = Path(str(identity.get("path", ""))).resolve()
    if (declared_path != expected_path.resolve()
            or identity.get("bytes") != actual["bytes"]
            or identity.get("sha256") != actual["sha256"]):
        raise AuditError(
            f"{label}: product-declared artifact identity is stale: "
            f"{identity!r} vs {actual!r}")
    return actual


def _projection_tuple(row: DetailRow) -> tuple[str, ...]:
    return tuple(_product_text(value) for value in row.product_projection())


def _inspect_projection_sheet(
        worksheet, rows: Sequence[DetailRow], *, side: str,
        snapshot: bool) -> dict[str, object]:
    physical = worksheet.iter_rows(values_only=True)
    header = tuple(next(physical, ()))
    expected_header = (
        ("Source row", "Route", *SHARED_HEADER, "Key (helper)")
        if snapshot else
        ("Comparison row", "Route", *SHARED_HEADER, "Key (helper)",
         "__CMP_E2_BUILD_FRESH_V1_C001_B_AJ"))
    if header != expected_header:
        raise AuditError(f"{worksheet.title}: projected header drift: {header!r}")
    digest = hashlib.sha256()
    expected_digest = hashlib.sha256()
    count = 0
    helper_tokens = set()
    for physical_row in physical:
        if not physical_row or all(value is None for value in physical_row):
            continue
        if len(physical_row) != len(expected_header):
            raise AuditError(f"{worksheet.title}: projected row width drift")
        if count >= len(rows):
            raise AuditError(f"{worksheet.title}: unexpected extra projected row")
        if snapshot and physical_row[0] != count + 1:
            raise AuditError(f"{worksheet.title}: source ordinal drift at {count + 1}")
        start, stop, helper_index = (1, -1, -1) if snapshot else (1, -2, -2)
        observed = tuple(_product_text(value)
                         for value in physical_row[start:stop])
        expected = _projection_tuple(rows[count])
        if observed != expected:
            raise AuditError(
                f"{worksheet.title}: projected source drift at row {count + 2}: "
                f"{observed!r} != {expected!r}")
        helper = str(physical_row[helper_index] or "")
        if not helper.startswith("__CMP_E2_KEY_V1_") or helper in helper_tokens:
            raise AuditError(f"{worksheet.title}: helper identity drift")
        helper_tokens.add(helper)
        digest.update(_canonical(observed))
        digest.update(b"\n")
        expected_digest.update(_canonical(expected))
        expected_digest.update(b"\n")
        count += 1
    if count != len(rows):
        raise AuditError(
            f"{worksheet.title}: projected rows {count} != {len(rows)}")
    return {
        "side": side, "rows": count, "columns": len(expected_header),
        "ordered_projection_sha256": digest.hexdigest(),
        "expected_ordered_projection_sha256": expected_digest.hexdigest(),
        "projection_exact": True, "helper_tokens_unique": True,
    }


def _source_only_counter(rows: Sequence[DetailRow]) -> Counter[tuple[object, ...]]:
    result: Counter[tuple[object, ...]] = Counter()
    for row in rows:
        claims = dict(row.source_only)
        result[(
            _product_text(row.route), _product_text(row.pm),
            _product_text(_project("ML Eff-Date", claims["MAIN_EFF_DATE"])),
            _product_text(claims["MAIN_ADT"]),
            _product_text(claims["CROSS_ADT"]),
        )] += 1
    return result


def _inspect_report_view(
        worksheet, *, expected_records: int, raw_tsn: Sequence[DetailRow] | None
        ) -> dict[str, object]:
    physical = worksheet.iter_rows(values_only=True)
    headers = [tuple(next(physical, ())) for _ in range(4)]
    if any(len(row) != 26 for row in headers):
        raise AuditError("Report View four-row header width drift")
    header_contract = (
        headers[0][24] == "TSN only"
        and headers[1][24] == "ML 2nd EFF"
        and headers[0][25] in (None, "")
        and headers[1][25] == "ML ADT"
        and headers[2][24] == "TSN only"
        and headers[2][25] in (None, "")
        and headers[3][24] in (None, "")
        and headers[3][25] == "CS ADT")
    if not header_contract:
        raise AuditError("Report View source-only header mapping drift")
    rows = [tuple(row) for row in physical
            if row and any(value is not None for value in row)]
    if len(rows) % 2:
        raise AuditError("Report View contains an orphan physical row")
    records = len(rows) // 2
    if records != expected_records:
        raise AuditError(
            f"Report View records {records} != expected {expected_records}")
    observed: Counter[tuple[object, ...]] = Counter()
    nonblank = Counter()
    for index in range(0, len(rows), 2):
        main, cross = rows[index:index + 2]
        if len(main) != 26 or len(cross) != 26:
            raise AuditError("Report View data-row width drift")
        if _product_text(main[2]) != _product_text(cross[2]):
            raise AuditError("Report View two-line Route identity drift")
        route = _product_text(main[2])
        pm = _product_text(main[5])
        ml2 = _product_text(main[24])
        adt = _product_text(main[25])
        cadt = _product_text(cross[25])
        nonblank["MAIN_EFF_DATE"] += int(bool(ml2))
        nonblank["MAIN_ADT"] += int(bool(adt))
        nonblank["CROSS_ADT"] += int(bool(cadt))
        if ml2 or adt or cadt:
            observed[(route, pm, ml2, adt, cadt)] += 1
    if raw_tsn is None:
        if any(nonblank.values()) or observed:
            raise AuditError(
                "normalized Report View unexpectedly recovered omitted source claims")
        expected_digest = _counter_digest(Counter())
    else:
        expected = _source_only_counter(raw_tsn)
        if observed != expected:
            raise AuditError(
                "raw Report View source-only mapping drift: "
                f"{_counter_difference_examples(observed, expected)!r}")
        expected_nonblank = len(raw_tsn)
        if dict(nonblank) != {
                "MAIN_EFF_DATE": expected_nonblank,
                "MAIN_ADT": expected_nonblank,
                "CROSS_ADT": expected_nonblank}:
            raise AuditError(
                f"raw Report View source-only nonblank census drift: {nonblank!r}")
        expected_digest = _counter_digest(expected)
    return {
        "records": records, "physical_data_rows": len(rows),
        "source_only_header_mapping_exact": header_contract,
        "source_only_nonblank_counts": dict(nonblank),
        "source_only_ledger_sha256": _counter_digest(observed),
        "expected_source_only_ledger_sha256": expected_digest,
        "source_only_mapping_exact": True,
    }


def _helper_outcome_exact(
        label: str, payload: dict[str, object],
        expected: dict[str, object]) -> dict[str, object]:
    result = payload.get("result")
    if not isinstance(result, dict):
        raise AuditError(f"product {label} outcome is missing")
    counts = result.get("counts")
    if not isinstance(counts, dict):
        raise AuditError(f"product {label} returned no structured counts")
    expected_counts = {"known": True, **EXPECTED_PRODUCT_COUNTS[label]}
    product_counts = dict(counts)
    product_counts["per_field_counts"] = {
        str(key).split(":", 1)[-1]: value
        for key, value in (counts.get("per_field_counts") or {}).items()
        if value
    }
    if product_counts != expected_counts:
        raise AuditError(
            f"product {label} returned-count drift: {product_counts!r}")
    if (result.get("status"), result.get("completion"), result.get("verdict"),
            result.get("skipped_inputs"), result.get("failed_inputs")) != (
                "ok", "complete", "diff", 0, 0):
        raise AuditError(f"product {label} outcome state drift: {result!r}")
    if result.get("warnings") or result.get("failures"):
        raise AuditError(f"product {label} reported warnings/failures")
    generation = result.get("artifact_generation")
    if not isinstance(generation, dict) or (
            generation.get("completion"), generation.get("publication_state"),
            generation.get("requested_mode")) != (
                "complete", "committed", "both"):
        raise AuditError(f"product {label} twin generation state drift")
    members = generation.get("members")
    flavors = sorted(
        str(member.get("flavor")) for member in members
        if isinstance(member, dict)) if isinstance(members, list) else []
    if flavors != ["formulas", "values"]:
        raise AuditError(f"product {label} twin manifest drift: {flavors!r}")
    return {
        "status": result.get("status"), "completion": result.get("completion"),
        "verdict": result.get("verdict"), "counts": product_counts,
        "artifact_generation": {
            "completion": generation.get("completion"),
            "publication_state": generation.get("publication_state"),
            "requested_mode": generation.get("requested_mode"),
            "flavors": flavors,
        },
        "matches_independent_expected_counts": True,
    }


def _expected_product_sheets(
        side_a: str, side_b: str, *, notes: bool, report_view: bool
        ) -> list[str]:
    names = [
        "Summary", "Spot Check", "Comparison", "Routes",
        f"Only in {side_a}", f"Only in {side_b}", side_a, side_b,
    ]
    if notes:
        names.append("Notes")
    if report_view:
        names.append("Report View")
    names.extend(["__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B"])
    return names


def _inspect_only_sheet(
        worksheet, *, present: str, missing: str,
        expected: Counter[tuple[object, ...]]) -> dict[str, object]:
    physical = worksheet.iter_rows(values_only=True)
    header = tuple(next(physical, ()))
    expected_header = (
        "Route", "PM", "#", f"{present} Row", f"Missing from {missing}",
        *ASSERTED_FIELDS)
    if header != expected_header:
        raise AuditError(f"{worksheet.title}: one-sided header drift: {header!r}")
    observed: Counter[tuple[object, ...]] = Counter()
    for row in physical:
        if not row or all(value is None for value in row):
            continue
        if len(row) != len(expected_header):
            raise AuditError(f"{worksheet.title}: one-sided row width drift")
        observed[(
            _product_text(row[0]), _product_text(row[1]),
            *(_product_text(value) for value in row[5:]),
        )] += 1
    if observed != expected:
        raise AuditError(
            f"{worksheet.title}: one-sided inventory drift: "
            f"{_counter_difference_examples(observed, expected)!r}")
    return {
        "rows": sum(observed.values()),
        "ledger_sha256": _counter_digest(observed),
        "expected_ledger_sha256": _counter_digest(expected),
        "inventory_exact": True,
    }


def _inspect_product_workbook(
        label: str, formulas_path: Path, values_path: Path,
        expected: dict[str, object], left: Sequence[DetailRow],
        right: Sequence[DetailRow], *, notes: bool,
        report_view: str | None, raw_tsn: Sequence[DetailRow]) -> dict[str, object]:
    side_a = str(expected["side_a"])
    side_b = str(expected["side_b"])
    expected_sheets = _expected_product_sheets(
        side_a, side_b, notes=notes, report_view=report_view is not None)
    formula_census = _formula_tag_census(formulas_path)
    value_formula_census = _formula_tag_census(values_path)
    expected_value_census = {
        "Summary": 12 if report_view is not None else 11,
        "Spot Check": 265,
        "Comparison": (
            expected["counts"]["paired_rows"]
            + expected["counts"]["paired_rows"]
            + expected["counts"]["side_a_only_rows"]
            + expected["counts"]["side_b_only_rows"]),
        "Routes": 0,
        f"Only in {side_a}": expected["counts"]["side_a_only_rows"],
        f"Only in {side_b}": expected["counts"]["side_b_only_rows"],
        side_a: 2 * len(left) + 1,
        side_b: 2 * len(right) + 1,
        "__CMP_E2_SNAPSHOT_A": 0, "__CMP_E2_SNAPSHOT_B": 0,
    }
    if notes:
        expected_value_census["Notes"] = 0
    if report_view is not None:
        expected_value_census["Report View"] = 0
    formula_only_sheets = {
        "Notes", "Report View", "__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B"}
    formula_contract = (
        set(formula_census) == set(expected_sheets)
        and value_formula_census == expected_value_census
        and sum(formula_census.values()) > sum(value_formula_census.values()) > 0
        and formula_census["Comparison"] > value_formula_census["Comparison"]
        and formula_census[side_a] == value_formula_census[side_a]
        and formula_census[side_b] == value_formula_census[side_b]
        and all(formula_census.get(sheet, 0) == 0
                for sheet in formula_only_sheets if sheet in formula_census))
    if not formula_contract:
        raise AuditError(
            f"product {label} formula/value flavor contract drift: "
            f"formulas={formula_census!r} values={value_formula_census!r} "
            f"expected_values={expected_value_census!r}")
    formula_book = load_workbook(
        formulas_path, read_only=True, data_only=False)
    try:
        if formula_book.sheetnames != expected_sheets:
            raise AuditError(
                f"product {label} formulas sheet universe drift: "
                f"{formula_book.sheetnames!r}")
    finally:
        formula_book.close()

    values = load_workbook(values_path, read_only=True, data_only=True)
    try:
        if values.sheetnames != expected_sheets:
            raise AuditError(
                f"product {label} values sheet universe drift: "
                f"{values.sheetnames!r}")
        if (values["__CMP_E2_SNAPSHOT_A"].sheet_state != "veryHidden"
                or values["__CMP_E2_SNAPSHOT_B"].sheet_state != "veryHidden"):
            raise AuditError(f"product {label} snapshots are not veryHidden")

        comparison = values["Comparison"]
        physical = comparison.iter_rows(values_only=True)
        header = tuple(next(physical, ()))
        state_header = "__CMP_E1_STATE_V1_C001_P0000_P0031"
        expected_header = (
            "Route", "PM", "#", f"{side_a} Row", f"{side_b} Row",
            "Status", "Diffs", *ASSERTED_FIELDS, state_header)
        if header != expected_header:
            raise AuditError(
                f"product {label} Comparison header drift: {header!r}")
        observed_pairs: Counter[tuple[object, ...]] = Counter()
        status_counts: Counter[str] = Counter()
        per_field: Counter[str] = Counter()
        differing_rows = differing_cells = union_rows = 0
        for row in physical:
            if not row or all(value is None for value in row):
                continue
            if len(row) != len(expected_header):
                raise AuditError(f"product {label} Comparison row width drift")
            union_rows += 1
            status = str(row[5])
            status_counts[status] += 1
            if status != "Both":
                continue
            state = str(row[-1])
            if len(state) != len(ASSERTED_FIELDS) or set(state) - {"E", "D"}:
                raise AuditError(f"product {label} paired state-vector drift")
            diffs = state.count("D")
            if row[6] != diffs:
                raise AuditError(f"product {label} Diffs/state disagreement")
            displays = tuple(_product_text(value) for value in row[7:-1])
            observed_pairs[(
                _product_text(row[0]), _product_text(row[1]), diffs,
                *displays, state,
            )] += 1
            differing_rows += int(diffs > 0)
            differing_cells += diffs
            for field, code in zip(ASSERTED_FIELDS, state):
                per_field[field] += int(code == "D")
        expected_pairs = expected["_paired"]
        if observed_pairs != expected_pairs:
            raise AuditError(
                f"product {label} paired cell ledger drift: "
                f"{_counter_difference_examples(observed_pairs, expected_pairs)!r}")
        expected_statuses = {
            "Both": expected["counts"]["paired_rows"],
            f"{side_a} only": expected["counts"]["side_a_only_rows"],
            f"{side_b} only": expected["counts"]["side_b_only_rows"],
        }
        expected_statuses = {
            key: value for key, value in expected_statuses.items()
            if value or key == "Both"}
        if dict(status_counts) != expected_statuses:
            raise AuditError(
                f"product {label} status census drift: {status_counts!r}")
        nonzero_per_field = {
            field: per_field[field] for field in ASSERTED_FIELDS
            if per_field[field]}
        if (union_rows != sum(expected_statuses.values())
                or differing_rows != expected["counts"]["differing_rows"]
                or differing_cells != expected["counts"]["differing_cells"]
                or nonzero_per_field != expected["counts"]["per_field_counts"]):
            raise AuditError(f"product {label} independently read counts drift")

        only_a = _inspect_only_sheet(
            values[f"Only in {side_a}"], present=side_a, missing=side_b,
            expected=expected["_side_a_only"])
        only_b = _inspect_only_sheet(
            values[f"Only in {side_b}"], present=side_b, missing=side_a,
            expected=expected["_side_b_only"])
        snapshot_a = _inspect_projection_sheet(
            values["__CMP_E2_SNAPSHOT_A"], left, side=side_a, snapshot=True)
        snapshot_b = _inspect_projection_sheet(
            values["__CMP_E2_SNAPSHOT_B"], right, side=side_b, snapshot=True)
        visible_a = _inspect_projection_sheet(
            values[side_a], left, side=side_a, snapshot=False)
        visible_b = _inspect_projection_sheet(
            values[side_b], right, side=side_b, snapshot=False)
        notes_contract = None
        if notes:
            note_text = "\n".join(
                str(value) for row in values["Notes"].iter_rows(values_only=True)
                for value in row if value is not None)
            notes_contract = {
                "declares_route_plus_postmile": (
                    "Rows are keyed on Route + Postmile (PM)." in note_text),
                "declares_county_identity": "County" in note_text,
                "declares_district_assertion": "District" in note_text,
            }
            if not notes_contract["declares_route_plus_postmile"]:
                raise AuditError(f"product {label} weak-key Notes declaration drift")
        report_view_result = None
        if report_view is not None:
            report_view_result = _inspect_report_view(
                values["Report View"],
                expected_records=union_rows,
                raw_tsn=raw_tsn if report_view == "raw" else None)
    finally:
        values.close()
    return {
        "formula_tag_census": formula_census,
        "values_formula_tag_census": value_formula_census,
        "formula_value_flavors_structurally_exact": True,
        "sheet_universe": expected_sheets,
        "counts": {
            "union_rows": union_rows, "paired_rows": status_counts["Both"],
            "side_a_only_rows": expected["counts"]["side_a_only_rows"],
            "side_b_only_rows": expected["counts"]["side_b_only_rows"],
            "differing_rows": differing_rows,
            "differing_cells": differing_cells,
            "per_field_counts": nonzero_per_field,
        },
        "paired_cell_ledger_sha256": _counter_digest(observed_pairs),
        "expected_paired_cell_ledger_sha256": expected[
            "paired_ledger_sha256"],
        "comparison_header": list(header),
        "comparison_identity_columns": ["Route", "PM"],
        "district_column_present": "District" in header,
        "county_column_present": "County" in header,
        "only_in": {side_a: only_a, side_b: only_b},
        "snapshots": {side_a: snapshot_a, side_b: snapshot_b},
        "visible_source_sheets": {side_a: visible_a, side_b: visible_b},
        "notes_contract": notes_contract,
        "report_view": report_view_result,
    }


def _loaded_product_manifest_current(
        manifest: dict[str, object]) -> dict[str, object]:
    entries = manifest.get("entries")
    if not isinstance(entries, list) or manifest.get("file_count") != len(entries):
        raise AuditError("product module manifest shape drift")
    canonical = json.dumps(
        entries, sort_keys=True, separators=(",", ":")).encode("utf-8")
    if hashlib.sha256(canonical).hexdigest() != manifest.get(
            "canonical_json_sha256"):
        raise AuditError("product module manifest canonical digest drift")
    detail = []
    for entry in entries:
        if not isinstance(entry, dict):
            raise AuditError("product module manifest entry drift")
        path = REPO_ROOT / "scripts" / str(entry.get("relative_path", ""))
        observed = ({"bytes": path.stat().st_size, "sha256": _sha_file(path)}
                    if path.is_file() else None)
        expected = {"bytes": entry.get("bytes"), "sha256": entry.get("sha256")}
        detail.append({
            "relative_path": entry.get("relative_path"),
            "expected": expected, "observed": observed,
            "current": observed == expected,
        })
    if not detail or not all(item["current"] for item in detail):
        raise AuditError("loaded product code changed after witness execution")
    return {
        "file_count": len(detail),
        "canonical_json_sha256": manifest.get("canonical_json_sha256"),
        "all_current": True, "entries": detail,
    }


def _run_self_gate() -> dict[str, object]:
    completed = subprocess.run(
        [sys.executable, str(SELF_GATE_PATH)], cwd=REPO_ROOT,
        text=True, capture_output=True, timeout=120, check=False)
    if completed.returncode != 0:
        raise AuditError(
            "Intersection Detail Stage-8 mutation gate failed: "
            f"{completed.stderr or completed.stdout}")
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    try:
        payload = json.loads(lines[-1])
    except (IndexError, json.JSONDecodeError) as exc:
        raise AuditError("Stage-8 mutation gate emitted no JSON result") from exc
    if payload != {"status": "pass", "assertions": 31}:
        raise AuditError(f"Stage-8 mutation gate result drift: {payload!r}")
    return {
        "status": "executed_pass", "assertions": 31,
        "stdout": completed.stdout.strip(),
        "gate": _file_identity(SELF_GATE_PATH),
    }


def _inspect_product(
        product_root: Path, expected: dict[str, dict[str, object]],
        rows_by_source: dict[str, Sequence[DetailRow]]) -> dict[str, object]:
    result_path = product_root / "product-witness-result.json"
    payload = _read_json(result_path)
    consolidations = payload.get("consolidations")
    comparisons = payload.get("comparisons")
    if not isinstance(consolidations, dict) or set(consolidations) != {"excel", "pdf"}:
        raise AuditError("product witness consolidation manifest drift")
    if not isinstance(comparisons, dict) or set(comparisons) != set(expected):
        raise AuditError("product witness comparison-leg universe drift")
    excel_path = product_root / "intersection_detail_excel_consolidated.xlsx"
    pdf_path = product_root / "intersection_detail_pdf_consolidated.xlsx"
    consolidation_rows = {
        "excel": rows_by_source["tsmis_excel"],
        "pdf": rows_by_source["tsmis_pdf"],
    }
    consolidation_paths = {"excel": excel_path, "pdf": pdf_path}
    consolidation_inspection = {}
    for flavor in ("excel", "pdf"):
        item = consolidations[flavor]
        if not isinstance(item, dict) or (
                item.get("status"), item.get("completion"),
                item.get("skipped_inputs"), item.get("failed_inputs")) != (
                    "ok", "complete", 0, 0):
            raise AuditError(f"product {flavor} consolidation outcome drift")
        _validate_file_identity(
            item.get("output", {}), consolidation_paths[flavor],
            f"product {flavor} consolidation")
        consolidation_inspection[flavor] = {
            **_inspect_consolidated(
                consolidation_paths[flavor], consolidation_rows[flavor]),
            "file_identity": _file_identity(consolidation_paths[flavor]),
        }

    leg_contract = {
        "excel_vs_tsn_raw": (
            rows_by_source["tsmis_excel"], rows_by_source["tsn_raw"],
            True, "raw"),
        "excel_vs_tsn_normalized": (
            rows_by_source["tsmis_excel"], rows_by_source["tsn_normalized"],
            True, "normalized"),
        "pdf_vs_tsn_raw": (
            rows_by_source["tsmis_pdf"], rows_by_source["tsn_raw"],
            True, None),
        "pdf_vs_tsn_normalized": (
            rows_by_source["tsmis_pdf"], rows_by_source["tsn_normalized"],
            True, None),
        "pdf_vs_excel": (
            rows_by_source["tsmis_pdf"], rows_by_source["tsmis_excel"],
            False, None),
    }
    inspected = {}
    for label, (left, right, notes, report_view) in leg_contract.items():
        item = comparisons[label]
        if not isinstance(item, dict):
            raise AuditError(f"product {label} witness payload drift")
        outcome = _helper_outcome_exact(label, item, expected[label])
        outputs = item.get("outputs")
        if not isinstance(outputs, dict) or set(outputs) != {"formulas", "values"}:
            raise AuditError(f"product {label} output manifest drift")
        formulas = product_root / f"{label}.xlsx"
        values = product_root / f"{label} (values).xlsx"
        _validate_file_identity(outputs["formulas"], formulas, f"{label} formulas")
        _validate_file_identity(outputs["values"], values, f"{label} values")
        inspected[label] = {
            "returned_outcome": outcome,
            "outputs": {
                "formulas": {"path": str(formulas.resolve()),
                             "bytes": formulas.stat().st_size,
                             "sha256": _sha_file(formulas)},
                "values": {"path": str(values.resolve()),
                           "bytes": values.stat().st_size,
                           "sha256": _sha_file(values)},
            },
            "independent_workbook_inspection": _inspect_product_workbook(
                label, formulas, values, expected[label], left, right,
                notes=notes, report_view=report_view,
                raw_tsn=rows_by_source["tsn_raw"]),
        }
    loaded = payload.get("loaded_product_code")
    if not isinstance(loaded, dict):
        raise AuditError("product witness omitted loaded-code manifest")
    return {
        "witness_result": {
            "path": str(result_path.resolve()),
            "bytes": result_path.stat().st_size,
            "sha256": _sha_file(result_path),
        },
        "consolidations": consolidation_inspection,
        "comparisons": inspected,
        "loaded_product_code": loaded,
        "loaded_product_code_current": _loaded_product_manifest_current(loaded),
    }


def _triangle(rows_by_source: dict[str, Sequence[DetailRow]]) -> dict[str, object]:
    key = ("108", "TUO", "", "5.87")
    result = {"physical_key": list(key), "sources": {}}
    for label, rows in rows_by_source.items():
        matches = [row for row in rows if row.physical_key == key]
        result["sources"][label] = [
            {
                "source_ref": row.source_ref,
                "district": row.district,
                "hg": row.values[ASSERTED_FIELDS.index("HG")],
                "description": row.values[ASSERTED_FIELDS.index("Description")],
            }
            for row in matches
        ]
    return result


def run(args: argparse.Namespace) -> dict[str, object]:
    mutation_gate = _run_self_gate()
    bindings = {
        "tsmis_excel": _bind_tree("tsmis_excel", args.tsmis_xlsx_root),
        "tsmis_pdf": _bind_tree("tsmis_pdf", args.tsmis_pdf_root),
        "tsn_raw": _bind_file("tsn_raw", args.tsn_raw),
        "tsn_pdf": _bind_file("tsn_pdf", args.tsn_pdf),
        "tsn_normalized": _bind_file("tsn_normalized", args.tsn_normalized),
    }
    origin_bindings = {
        "tsmis_excel": _bind_tree(
            "tsmis_excel", args.origin_tsmis_xlsx_root),
        "tsmis_pdf": _bind_tree("tsmis_pdf", args.origin_tsmis_pdf_root),
        "tsn_raw": _bind_file("tsn_raw", args.origin_tsn_raw),
        "tsn_pdf": _bind_file("tsn_pdf", args.origin_tsn_pdf),
    }
    dependencies = _accepted_dependencies(args)
    excel = _parse_tsmis_excel(args.tsmis_xlsx_root)
    pdf = _parse_tsmis_pdf(args.tsmis_pdf_root)
    tsn = _parse_tsn_raw(args.tsn_raw)
    normalized = _parse_tsn_normalized(args.tsn_normalized)
    excel_rows = excel["rows_data"]
    pdf_rows = pdf["rows_data"]
    tsn_rows = tsn["rows_data"]
    normalized_rows = normalized["rows_data"]
    comparisons = {
        "excel_vs_tsn_raw": _comparison(
            "TSMIS Excel vs raw TSN", excel_rows, tsn_rows),
        "pdf_vs_tsn_raw": _comparison(
            "TSMIS PDF vs raw TSN", pdf_rows, tsn_rows),
        "pdf_vs_excel": _comparison(
            "TSMIS PDF vs TSMIS Excel", pdf_rows, excel_rows),
        "raw_vs_normalized": _comparison(
            "raw TSN vs normalized r7", tsn_rows, normalized_rows),
    }
    for label, expected_counts in EXPECTED_SOURCE_COUNTS.items():
        if comparisons[label]["counts"] != expected_counts:
            raise AuditError(
                f"{label}: full-physical source counts drifted: "
                f"{comparisons[label]['counts']!r}")

    rows_by_source = {
        "tsmis_excel": excel_rows, "tsmis_pdf": pdf_rows,
        "tsn_raw": tsn_rows, "tsn_normalized": normalized_rows,
    }
    product_expected = {
        "excel_vs_tsn_raw": _product_expected(
            "excel_vs_tsn_raw", "TSMIS", "TSN", excel_rows, tsn_rows),
        "excel_vs_tsn_normalized": _product_expected(
            "excel_vs_tsn_normalized", "TSMIS", "TSN",
            excel_rows, normalized_rows),
        "pdf_vs_tsn_raw": _product_expected(
            "pdf_vs_tsn_raw", "TSMIS (PDF)", "TSN", pdf_rows, tsn_rows),
        "pdf_vs_tsn_normalized": _product_expected(
            "pdf_vs_tsn_normalized", "TSMIS (PDF)", "TSN",
            pdf_rows, normalized_rows),
        "pdf_vs_excel": _product_expected(
            "pdf_vs_excel", "TSMIS (PDF)", "TSMIS (Excel)",
            pdf_rows, excel_rows),
    }
    production = _inspect_product(
        args.product_root, product_expected, rows_by_source)

    source_invariants = {
        "all_authoritative_source_members_match_frozen_bindings": all(
            all(capture["observed"].get(key) == value
                for key, value in capture["binding"].items()
                if key != "suffix")
            for capture in bindings.values()),
        "private_capture_matches_live_authoritative_origins": (
            all(origin_bindings[label]["observed"] == bindings[label]["observed"]
                for label in ("tsmis_excel", "tsmis_pdf", "tsn_raw", "tsn_pdf"))),
        "all_accepted_dependencies_exact": all(
            all(item["checks"].values())
            for item in (
                dependencies["normalized_outcome"],
                dependencies["stage6_raw_to_normalized"],
                dependencies["tsn_xlsx_to_pdf"])),
        "all_four_source_row_counts_exact": (
            len(excel_rows), len(pdf_rows), len(tsn_rows),
            len(normalized_rows)) == (16_459, 16_459, 16_626, 16_626),
        "tsmis_pdf_all_1844_pages_16459_pairs_zero_residue": (
            pdf["pdf_reconciliation"] == {
                "pages": 1_844, "row_a": 16_459, "row_b": 16_459,
                "orphans": 0, "old_layout_rows": 0,
                "row_a_bands": 8_285, "row_b_bands": 8_285}),
        "all_four_full_physical_source_comparisons_exact": all(
            comparisons[label]["counts"] == expected
            for label, expected in EXPECTED_SOURCE_COUNTS.items()),
        "raw_and_normalized_asserted_projection_exact": (
            comparisons["raw_vs_normalized"]["counts"]
            == EXPECTED_SOURCE_COUNTS["raw_vs_normalized"]),
        "raw_source_only_claims_all_16626_nonblank": (
            tsn["source_only_nonblank_counts"] == {
                "MAIN_EFF_DATE": 16_626, "MAIN_ADT": 16_626,
                "CROSS_ADT": 16_626}),
        "normalized_source_only_claim_columns_absent": (
            normalized["source_only_columns_present"] == []),
        "approved_physical_identity_census_exact": (
            tsn["identity"]["physical_identity"] == {
                "unique": 16_611, "duplicate_groups": 15,
                "duplicate_occurrences": 30, "max_multiplicity": 2}
            and tsn["identity"]["route_plus_numeric_pm_cross_county"] == {
                "keys": 78, "county_identities": 156}
            and tsn["identity"][
                "route_plus_complete_pp_plus_pm_cross_county"] == {
                    "keys": 71, "county_identities": 142}
            and len(tsn["identity"][
                "within_county_route_pm_complete_pp_collisions"]) == 6),
        "current_tsmis_pdf_excel_triangle_exact_nine_cells": (
            comparisons["pdf_vs_excel"]["counts"]["per_field_counts"]
            == {"Description": 8, "HG": 1}),
        "current_route108_hg_source_triangle_bound": (
            [item["hg"] for item in _triangle(rows_by_source)["sources"][
                "tsmis_excel"]] == ["U"]
            and [item["hg"] for item in _triangle(rows_by_source)["sources"][
                "tsmis_pdf"]] == ["D"]
            and [item["hg"] for item in _triangle(rows_by_source)["sources"][
                "tsn_raw"]] == ["D"]
            and [item["hg"] for item in _triangle(rows_by_source)["sources"][
                "tsn_normalized"]] == ["D"]),
    }
    product_books = {
        label: item["independent_workbook_inspection"]
        for label, item in production["comparisons"].items()
    }
    production_invariants = {
        "permanent_mutation_gate_executed_pass": (
            mutation_gate["status"] == "executed_pass"),
        "production_excel_pdf_consolidations_source_exact": all(
            item["projection_exact"]
            for item in production["consolidations"].values()),
        "production_all_five_formula_value_legs_independently_exact": (
            set(product_books) == set(EXPECTED_PRODUCT_COUNTS)
            and all(item["formula_value_flavors_structurally_exact"]
                    for item in product_books.values())),
        "production_all_five_paired_cell_ledgers_source_exact": all(
            item["paired_cell_ledger_sha256"]
            == item["expected_paired_cell_ledger_sha256"]
            for item in product_books.values()),
        "production_all_twenty_source_views_and_snapshots_source_exact": all(
            all(view["projection_exact"] for view in (
                *item["snapshots"].values(),
                *item["visible_source_sheets"].values()))
            for item in product_books.values()),
        "production_all_one_sided_inventories_source_exact": all(
            all(side["inventory_exact"] for side in item["only_in"].values())
            for item in product_books.values()),
        "production_raw_report_view_maps_all_source_only_claims": (
            product_books["excel_vs_tsn_raw"]["report_view"][
                "source_only_mapping_exact"] is True
            and product_books["excel_vs_tsn_raw"]["report_view"][
                "source_only_nonblank_counts"] == {
                    "MAIN_EFF_DATE": 16_626, "MAIN_ADT": 16_626,
                    "CROSS_ADT": 16_626}),
        "production_normalized_report_view_reproduces_three_omissions": (
            product_books["excel_vs_tsn_normalized"]["report_view"][
                "source_only_nonblank_counts"] == {
                    "MAIN_EFF_DATE": 0, "MAIN_ADT": 0, "CROSS_ADT": 0}),
        "production_pdf_tsn_report_view_absence_reproduced": (
            product_books["pdf_vs_tsn_raw"]["report_view"] is None
            and product_books["pdf_vs_tsn_normalized"]["report_view"] is None),
        "production_weak_route_pm_identity_and_source_visibility_gap_explicit": all(
            item["comparison_identity_columns"] == ["Route", "PM"]
            and not item["district_column_present"]
            and not item["county_column_present"]
            for item in product_books.values()),
        "production_raw_and_normalized_comparison_ledgers_identical": (
            product_books["excel_vs_tsn_raw"]["paired_cell_ledger_sha256"]
            == product_books["excel_vs_tsn_normalized"][
                "paired_cell_ledger_sha256"]
            and product_books["pdf_vs_tsn_raw"]["paired_cell_ledger_sha256"]
            == product_books["pdf_vs_tsn_normalized"][
                "paired_cell_ledger_sha256"]),
        "loaded_product_code_current_at_result_build": production[
            "loaded_product_code_current"]["all_current"],
    }
    audit_invariants = {**source_invariants, **production_invariants}
    source_truth_exact = all(source_invariants.values())
    stage8_complete = all(audit_invariants.values())
    triangle = _triangle(rows_by_source)
    findings = {
        "oracle_blocking": [],
        "source_export_deltas": [{
            "classification": "current TSMIS Excel export defect",
            "fact": (
                "At physical identity 108/TUO/<blank>/5.87, current TSMIS "
                "Excel says HG U while current TSMIS PDF plus raw and normalized "
                "TSN say HG D. The comparison correctly surfaces this cell."),
            "evidence": triangle,
        }, {
            "classification": "render-medium difference, not data loss",
            "fact": (
                "Eight TSMIS Excel Description cells contain trailing tab data "
                "that PDF cannot render; with the HG cell this is the exact "
                "nine-cell PDF-vs-Excel result."),
            "evidence": comparisons["pdf_vs_excel"][
                "difference_examples_first_100"],
        }],
        "product_red": [{
            "finding": "CMP-AUD-045",
            "fact": (
                "Every Intersection Detail product leg still keys on Route+PM "
                "and omits County although the approved identity is base Route, "
                "County, complete PP, numeric PM; raw TSN has 78 weak keys "
                "spanning 156 county identities."),
            "evidence": tsn["identity"],
        }, {
            "finding": "CMP-AUD-068",
            "fact": (
                "Both PDF-vs-TSN product legs omit the Report View that the "
                "Excel-vs-TSN legs produce."),
            "evidence": {
                label: item["sheet_universe"]
                for label, item in product_books.items()},
        }, {
            "finding": "CMP-AUD-070",
            "fact": (
                "Consolidation preserves every explicit member Route and physical "
                "S source claim exactly, but comparison re-derives Route/Suffix "
                "from Location and exposes neither claim."),
            "evidence": {
                "excel": production["consolidations"]["excel"],
                "pdf": production["consolidations"]["pdf"],
                "excel_source_claim_sha256": excel[
                    "explicit_tsmis_claims_sha256"],
                "pdf_source_claim_sha256": pdf[
                    "explicit_tsmis_claims_sha256"],
            },
        }, {
            "finding": "CMP-AUD-133",
            "fact": (
                "The raw product Report View maps all 16,626 MAIN_EFF_DATE, "
                "MAIN_ADT, and CROSS_ADT claims; the accepted normalized path "
                "renders all three columns blank. District/County sidecars also "
                "remain outside the visible comparison projection."),
            "evidence": {
                "raw": product_books["excel_vs_tsn_raw"]["report_view"],
                "normalized": product_books[
                    "excel_vs_tsn_normalized"]["report_view"],
            },
        }],
    }
    return {
        "schema_version": 1,
        "audit": "Stage 8 Intersection Detail authoritative four-source oracle",
        "status": "complete" if stage8_complete else "incomplete",
        "methodology": {
            "authority": (
                "Exact current 217-route TSMIS Excel/PDF pair, raw TSN XLSX "
                "and statewide print, accepted r7 normalization, accepted "
                "Stage-6 conservation, and accepted TSN XLSX/PDF mapping."),
            "independence": (
                "Source truth imports no application parser, comparator, schema, "
                "consolidator, or writer. Production ran in an isolated child; "
                "this oracle independently read every emitted workbook."),
            "physical_identity": [
                "base Route", "County", "complete PP", "numeric Post Mile"],
            "product_identity_observed": ["Route", "PM"],
            "source_asserted_fields": list(SOURCE_ASSERTED_FIELDS),
            "product_asserted_fields": list(ASSERTED_FIELDS),
        },
        "source_bindings": bindings,
        "live_origin_bindings": origin_bindings,
        "accepted_dependencies": dependencies,
        "sources": {
            "tsmis_excel": _public_source(excel),
            "tsmis_pdf": _public_source(pdf),
            "tsn_raw": _public_source(tsn),
            "tsn_normalized": _public_source(normalized),
        },
        "comparisons": comparisons,
        "source_triangle_route_108_pm_5_870": triangle,
        "product_expected_from_source": {
            label: _public_product_expected(item)
            for label, item in product_expected.items()},
        "production": production,
        "dependency_gates": {
            "stage8_intersection_detail_mutations": mutation_gate},
        "findings": findings,
        "audit_invariants": audit_invariants,
        "source_truth_exact": source_truth_exact,
        "production_tsmis_projection_exact": all(
            item["projection_exact"]
            for item in production["consolidations"].values()),
        "production_overlapping_comparison_cells_exact": all(
            item["paired_cell_ledger_sha256"]
            == item["expected_paired_cell_ledger_sha256"]
            for item in product_books.values()),
        "production_value_projection_exact": False,
        "production_comparison_semantics_exact": False,
        "stage8_base_oracle_complete": stage8_complete,
        "comparison_end_to_end_perfect": False,
        "provenance": {
            "code_identities": {
                "generator": _file_identity(GENERATOR_PATH),
                "product_helper": _file_identity(PRODUCT_HELPER_PATH),
                "self_gate": (_file_identity(SELF_GATE_PATH)
                              if SELF_GATE_PATH.is_file() else None),
            },
            "truth_dependency_boundary": [
                "build/phase3_xlsx_stream.py",
                "build/phase3_independent_oracle.py",
                "build/phase3_intersection_detail_oracle.py",
                "pdfplumber word extraction",
            ],
            "application_modules_imported": [],
        },
    }


def _atomic_write(path: Path, payload: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(payload, encoding="utf-8", newline="\n")
    temporary.replace(path)


def _unlink_if_present(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass


def _publication_current(
        args: argparse.Namespace, result: dict[str, object]
        ) -> tuple[bool, dict[str, object]]:
    observed_sources = {
        "tsmis_excel": _bind_tree("tsmis_excel", args.tsmis_xlsx_root),
        "tsmis_pdf": _bind_tree("tsmis_pdf", args.tsmis_pdf_root),
        "tsn_raw": _bind_file("tsn_raw", args.tsn_raw),
        "tsn_pdf": _bind_file("tsn_pdf", args.tsn_pdf),
        "tsn_normalized": _bind_file("tsn_normalized", args.tsn_normalized),
    }
    observed_origins = {
        "tsmis_excel": _bind_tree(
            "tsmis_excel", args.origin_tsmis_xlsx_root),
        "tsmis_pdf": _bind_tree("tsmis_pdf", args.origin_tsmis_pdf_root),
        "tsn_raw": _bind_file("tsn_raw", args.origin_tsn_raw),
        "tsn_pdf": _bind_file("tsn_pdf", args.origin_tsn_pdf),
    }
    dependency_current = _accepted_dependencies(args)
    source_checks = {
        label: observed_sources[label]["observed"]
        == result["source_bindings"][label]["observed"]
        for label in observed_sources}
    origin_checks = {
        label: observed_origins[label]["observed"]
        == result["live_origin_bindings"][label]["observed"]
        for label in observed_origins}
    dependency_checks = {
        label: dependency_current["identities"][label]["observed"]
        == result["accepted_dependencies"]["identities"][label]["observed"]
        for label in dependency_current["identities"]}
    code_checks = {}
    for label, path in (
            ("generator", GENERATOR_PATH),
            ("product_helper", PRODUCT_HELPER_PATH),
            ("self_gate", SELF_GATE_PATH)):
        current = _file_identity(path)
        code_checks[label] = (
            current == result["provenance"]["code_identities"][label])

    artifact_checks = {}
    witness_expected = result["production"]["witness_result"]
    witness_current = _file_identity(
        Path(str(witness_expected["path"])))
    artifact_checks["witness_result"] = witness_current == witness_expected
    for label, consolidation in result["production"]["consolidations"].items():
        expected = consolidation["file_identity"]
        current = _file_identity(Path(str(expected["path"])))
        artifact_checks[f"consolidation:{label}"] = current == expected
    for label, comparison in result["production"]["comparisons"].items():
        for flavor, expected in comparison["outputs"].items():
            current = _file_identity(Path(str(expected["path"])))
            artifact_checks[f"{label}:{flavor}"] = current == expected
    loaded_product = _loaded_product_manifest_current(
        result["production"]["loaded_product_code"])
    flags = [
        *source_checks.values(), *origin_checks.values(),
        *dependency_checks.values(), *code_checks.values(),
        *artifact_checks.values(), loaded_product["all_current"],
    ]
    return all(flags), {
        "source_checks": source_checks, "origin_checks": origin_checks,
        "dependency_checks": dependency_checks, "code_checks": code_checks,
        "artifact_checks": artifact_checks,
        "loaded_product_code_current": loaded_product["all_current"],
    }


def _write_decision(
        path: Path, output: Path, result: dict[str, object], *,
        accepted: bool, reason: str, postwrite_current: bool,
        postwrite_detail: dict[str, object], open_findings_authorized: bool
        ) -> dict[str, object]:
    identity = _file_identity(output)
    decision = {
        "schema_version": 1, "accepted": accepted, "reason": reason,
        "audit": result.get("audit"), "result": str(output.resolve()),
        "result_bytes": identity["bytes"],
        "result_sha256": identity["sha256"],
        "source_truth_exact": result.get("source_truth_exact", False),
        "production_tsmis_projection_exact": result.get(
            "production_tsmis_projection_exact", False),
        "production_overlapping_comparison_cells_exact": result.get(
            "production_overlapping_comparison_cells_exact", False),
        "production_value_projection_exact": result.get(
            "production_value_projection_exact", False),
        "production_comparison_semantics_exact": result.get(
            "production_comparison_semantics_exact", False),
        "stage8_base_oracle_complete": result.get(
            "stage8_base_oracle_complete", False),
        "comparison_end_to_end_perfect": result.get(
            "comparison_end_to_end_perfect", False),
        "open_product_findings_authorized": open_findings_authorized,
        "post_result_write_revalidation": postwrite_current,
        "post_result_write_identities": postwrite_detail,
    }
    _atomic_write(path, json.dumps(
        decision, ensure_ascii=False, sort_keys=True, indent=2) + "\n")
    return decision


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tsmis-xlsx-root", type=Path,
                        default=DEFAULT_TSMIS_XLSX_ROOT)
    parser.add_argument("--tsmis-pdf-root", type=Path,
                        default=DEFAULT_TSMIS_PDF_ROOT)
    parser.add_argument("--tsn-raw", type=Path, default=DEFAULT_TSN_RAW)
    parser.add_argument("--tsn-pdf", type=Path, default=DEFAULT_TSN_PDF)
    parser.add_argument("--tsn-normalized", type=Path,
                        default=DEFAULT_TSN_NORMALIZED)
    parser.add_argument("--tsn-normalized-sidecar", type=Path,
                        default=DEFAULT_TSN_NORMALIZED_SIDECAR)
    parser.add_argument("--stage6-result", type=Path,
                        default=DEFAULT_STAGE6_RESULT)
    parser.add_argument("--stage6-acceptance", type=Path,
                        default=DEFAULT_STAGE6_ACCEPTANCE)
    parser.add_argument("--tsn-cross-format", type=Path,
                        default=DEFAULT_TSN_CROSS_FORMAT)
    parser.add_argument("--product-root", type=Path,
                        default=DEFAULT_PRODUCT_ROOT)
    parser.add_argument("--origin-tsmis-xlsx-root", type=Path,
                        default=DEFAULT_TSMIS_XLSX_ROOT)
    parser.add_argument("--origin-tsmis-pdf-root", type=Path,
                        default=DEFAULT_TSMIS_PDF_ROOT)
    parser.add_argument("--origin-tsn-raw", type=Path,
                        default=DEFAULT_TSN_RAW)
    parser.add_argument("--origin-tsn-pdf", type=Path,
                        default=DEFAULT_TSN_PDF)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--allow-open-findings", action="store_true",
        help=(
            "accept the completed source/product witness while the exact "
            "documented product findings remain open"))
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    acceptance_path = args.output.with_suffix(
        args.output.suffix + ".acceptance.json")
    rejection_path = args.output.with_suffix(
        args.output.suffix + ".rejection.json")
    _unlink_if_present(acceptance_path)
    _unlink_if_present(rejection_path)
    try:
        result = run(args)
    except Exception as exc:
        result = {
            "schema_version": 1,
            "audit": "Stage 8 Intersection Detail authoritative four-source oracle",
            "status": "failed",
            "error": {"type": type(exc).__name__, "message": str(exc)},
            "source_truth_exact": False,
            "production_tsmis_projection_exact": False,
            "production_overlapping_comparison_cells_exact": False,
            "production_value_projection_exact": False,
            "production_comparison_semantics_exact": False,
            "stage8_base_oracle_complete": False,
            "comparison_end_to_end_perfect": False,
        }
        _atomic_write(args.output, json.dumps(
            result, ensure_ascii=False, sort_keys=True, indent=2) + "\n")
        decision = _write_decision(
            rejection_path, args.output, result, accepted=False,
            reason="oracle_execution_failed", postwrite_current=False,
            postwrite_detail={}, open_findings_authorized=False)
        print(json.dumps({
            "accepted": False, "reason": decision["reason"],
            "error": result["error"], "output": str(args.output.resolve()),
            "rejection": str(rejection_path.resolve()),
        }, ensure_ascii=False, separators=(",", ":")))
        return 1

    _atomic_write(args.output, json.dumps(
        result, ensure_ascii=False, sort_keys=True, indent=2, default=str) + "\n")
    postwrite_current, postwrite_detail = _publication_current(args, result)
    complete = result["stage8_base_oracle_complete"] is True
    open_findings = bool(result.get("findings", {}).get("product_red"))
    accepted = bool(
        complete and postwrite_current
        and (not open_findings or args.allow_open_findings))
    if not complete:
        reason = "stage8_base_oracle_incomplete"
    elif not postwrite_current:
        reason = "post_result_write_revalidation_failed"
    elif open_findings and not args.allow_open_findings:
        reason = "open_product_findings_require_explicit_authorization"
    else:
        reason = "accepted_complete_audit_with_documented_open_product_findings"
    decision_path = acceptance_path if accepted else rejection_path
    decision = _write_decision(
        decision_path, args.output, result, accepted=accepted, reason=reason,
        postwrite_current=postwrite_current,
        postwrite_detail=postwrite_detail,
        open_findings_authorized=bool(args.allow_open_findings))
    print(json.dumps({
        "status": result["status"], "accepted": decision["accepted"],
        "reason": decision["reason"],
        "output": str(args.output.resolve()),
        "bytes": args.output.stat().st_size,
        "sha256": _sha_file(args.output),
        "decision": str(decision_path.resolve()),
        "source_rows": {
            label: source["rows"] for label, source in result["sources"].items()
        },
        "comparison_counts": {
            label: comparison["counts"]
            for label, comparison in result["comparisons"].items()
        },
    }, ensure_ascii=False, separators=(",", ":")))
    return 0 if accepted else 1


if __name__ == "__main__":
    raise SystemExit(main())
