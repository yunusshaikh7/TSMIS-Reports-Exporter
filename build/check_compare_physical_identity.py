"""Phase-4 L0a gate: physical identity is structured, county-aware, and lossless.

This check is deliberately small and hermetic.  Real TSN collision counts are bound in
``docs/planning/comparison-perfection/comparison-phase4-red-fixture-index.md``; this file locks the engine and
adapter mechanisms that must make those collisions impossible to mis-pair.

CMP-AUD-045.  This file is BOTH a green contract and the finding's live red fixture.
``TESTS`` are the typed identity core, which is green.  ``KNOWN_RED`` are the four
family-projector contracts that still fail, because the projectors hand the engine
plain route/PM strings instead of a ``PhysicalKey``.  Each known-red contract asserts
that the defect is still present AND still has its recorded signature, so the check
fails if the defect drifts or is fixed — it is never a gate you learn to ignore.
Capture it red before changing loader semantics; promote KNOWN_RED into TESTS when the
projector batch lands.
"""
from pathlib import Path
import sys
from tempfile import TemporaryDirectory


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from compare_core import (  # noqa: E402
    CompareSchema,
    count_diffs,
    keys_for,
    pair_occurrences_by_similarity,
    run_compare,
    union_keys,
)
import comparison_contract as cc  # noqa: E402
from events import Events  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
import compare_env  # noqa: E402
import compare_highway_sequence_pdf as hslpdf  # noqa: E402
import compare_highway_sequence_tsn as hsl  # noqa: E402
import compare_intersection_detail_tsn as idt  # noqa: E402
import compare_ramp_detail_pdf as rdpdf  # noqa: E402
import compare_ramp_detail_tsn as rd  # noqa: E402

def _counts(schema, rows_a, rows_b):
    keys_a = keys_for(rows_a, True, schema.key_field, schema.key_normalizer)
    keys_b = keys_for(rows_b, True, schema.key_field, schema.key_normalizer)
    pairing = pair_occurrences_by_similarity(
        schema, rows_a, rows_b, keys_a, keys_b, True, Events())
    union = union_keys(pairing.keys_a, pairing.keys_b)
    counts = count_diffs(
        schema, rows_a, rows_b, pairing.keys_a, pairing.keys_b, union, True)
    return pairing, counts


def _assert_exact_physical_swap(schema, rows_a, rows_b, expected_cells,
                                expected_postmile):
    pairing, counts = _counts(schema, rows_a, rows_b)
    assert all(type(key[1]) is cc.PhysicalKey
               for key in pairing.keys_a + pairing.keys_b), (
                   "environment projector did not supply PhysicalKey", pairing)
    assert counts["both"] == 2, counts
    assert counts["t_only"] == counts["n_only"] == 0, counts
    assert counts["diff_rows"] == 2, counts
    assert counts["diff_cells"] == expected_cells, counts
    assert pairing.pairing_trace == (), (
        "distinct physical locations must not enter duplicate-key pairing",
        pairing.pairing_trace,
    )
    identities = {(key[0], key[1]) for key in pairing.keys_a}
    assert len(identities) == 2, pairing.keys_a
    assert {
        dict(key[1].physical_identity.canonical_components)["postmile"]
        for key in pairing.keys_a + pairing.keys_b
    } == {expected_postmile}


def _engine_claims(side, ordinal):
    return (
        cc.RawIdentityClaim("source_side", side),
        cc.RawIdentityClaim("source_ordinal", ordinal),
    )


def _engine_key(side, ordinal, source_display):
    identity = cc.make_physical_identity(
        "001", "ORA", "R1.000", _engine_claims(side, ordinal),
        f"{side}-specific identity display")
    return cc.physical_key(source_display, identity)


def _engine_rows():
    # PM is deliberately the SECOND schema field. A whole-row scan could pick
    # the decoy string in Payload and falsely bless the wrong key attachment.
    rows_a = [
        ["001", "decoy-one", _engine_key("A", 1, "A raw PM one"), "alpha"],
        ["001", "decoy-two", _engine_key("A", 2, "A raw PM two"), "beta"],
    ]
    rows_b = [
        ["001", "decoy-two", _engine_key("B", 1, "B raw PM one"), "beta"],
        ["001", "decoy-one", _engine_key("B", 2, "B raw PM two"), "alpha"],
    ]
    return rows_a, rows_b


_ENGINE_SCHEMA = CompareSchema(
    report_name="identity-probe",
    header=["Payload", "PM", "Value"],
    side_a="A",
    side_b="B",
    key_field=1,
)


def test_engine_preserves_each_side_raw_claim():
    """Occurrence renumbering retains each real PhysicalKey and exact claims."""
    rows_a, rows_b = _engine_rows()
    pairing, counts = _counts(_ENGINE_SCHEMA, rows_a, rows_b)
    assert counts["diff_cells"] == 0, counts
    assert all(type(key[1]) is cc.PhysicalKey
               for key in pairing.keys_a + pairing.keys_b)
    assert [key[1].physical_identity.raw_claims for key in pairing.keys_a] == [
        _engine_claims("A", 1), _engine_claims("A", 2)]
    assert [key[1].physical_identity.raw_claims for key in pairing.keys_b] == [
        _engine_claims("B", 1), _engine_claims("B", 2)], (
            "pairing replaced side-B raw identity claims", pairing.keys_b)
    assert pairing.pairing_trace[0].key_components == (
        "001", "ORA", "R1.000")
    assert pairing.keys_a[0][1] is rows_a[0][2]
    assert pairing.keys_b[0][1] is rows_b[0][2]


def test_loaded_side_roundtrip_and_claim_domain():
    rows_a, _rows_b = _engine_rows()
    claims = (
        cc.RawIdentityClaim("none", None),
        cc.RawIdentityClaim("blank", ""),
        cc.RawIdentityClaim("bool", True),
        cc.RawIdentityClaim("int", 7),
        cc.RawIdentityClaim("float", 1.25),
        cc.RawIdentityClaim("negative_zero", -0.0),
        cc.RawIdentityClaim("text", " exact "),
    )
    identity = cc.make_physical_identity(
        "001", "ORA", "R1.000", claims, "ignored side display")
    rows_a[0][2] = cc.physical_key("source cell text", identity)
    loaded = cc.LoadedSide(rows=tuple(tuple(row) for row in rows_a),
                           completion="partial")
    restored = cc.from_json(cc.to_json(loaded))
    restored_key = restored.rows[0][2]
    assert type(restored_key) is cc.PhysicalKey
    assert str(restored_key) == "source cell text"
    assert restored_key.physical_identity == identity
    assert restored_key.physical_identity.raw_claims == claims
    assert cc.RawIdentityClaim("typed", True) != cc.RawIdentityClaim("typed", 1)
    assert cc.RawIdentityClaim("typed", 1) != cc.RawIdentityClaim("typed", 1.0)
    assert cc.RawIdentityClaim("typed", -0.0) != cc.RawIdentityClaim("typed", 0.0)
    for invalid in ([], {}, ("tuple",), float("nan"), float("inf"), object()):
        try:
            cc.RawIdentityClaim("invalid", invalid)
        except ValueError:
            pass
        else:
            raise AssertionError(("raw claim accepted non-JSON scalar", invalid))


def test_route_authority_and_uniform_typed_mode():
    rows_a, rows_b = _engine_rows()
    expected = keys_for(rows_a, True, _ENGINE_SCHEMA.key_field)
    assert expected[0][0] == "001"
    assert expected[0][1] is rows_a[0][2]
    route_less = [["decoy", rows_a[0][2], "alpha"]]
    route_less_keys = keys_for(
        route_less, False, _ENGINE_SCHEMA.key_field)
    assert route_less_keys == [("001", rows_a[0][2], 1)]

    mismatched = [list(rows_a[0])]
    mismatched[0][0] = "002"
    try:
        keys_for(mismatched, True, _ENGINE_SCHEMA.key_field)
    except ValueError as exc:
        assert "outer route" in str(exc)
    else:
        raise AssertionError("outer/canonical route mismatch was accepted")

    type_identity = cc.make_physical_identity(
        "1", "ORA", "R1.000", _engine_claims("typed-route", 1),
        "ignored")
    type_mismatch = [[
        1, "decoy", cc.physical_key("raw", type_identity), "alpha"]]
    try:
        keys_for(type_mismatch, True, _ENGINE_SCHEMA.key_field)
    except ValueError as exc:
        assert "outer route" in str(exc)
    else:
        raise AssertionError("coercible non-string outer route was accepted")

    keys_a = keys_for(rows_a, True, _ENGINE_SCHEMA.key_field)
    legacy_b = [list(row) for row in rows_b]
    legacy_b[0][2] = "legacy PM one"
    legacy_b[1][2] = "legacy PM two"
    keys_b = keys_for(legacy_b, True, _ENGINE_SCHEMA.key_field)
    try:
        pair_occurrences_by_similarity(
            _ENGINE_SCHEMA, rows_a, legacy_b, keys_a, keys_b, True, Events())
    except ValueError as exc:
        assert "every key" in str(exc)
    else:
        raise AssertionError("mixed typed/legacy key mode was accepted")

    within_side_mixed = [list(row) for row in rows_b]
    within_side_mixed[0][2] = "legacy PM"
    within_keys = keys_for(
        within_side_mixed, True, _ENGINE_SCHEMA.key_field)
    try:
        pair_occurrences_by_similarity(
            _ENGINE_SCHEMA, rows_a, within_side_mixed,
            keys_a, within_keys, True, Events())
    except ValueError as exc:
        assert "every key" in str(exc)
    else:
        raise AssertionError("within-side mixed key mode was accepted")

    # The all-legacy path retains the exact key-field behavior.
    legacy = [["001", "not-the-key", "PM bytes", "value"]]
    assert keys_for(legacy, True, 1) == [("001", "PM bytes", 1)]

    with TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        mismatch_result = run_compare(
            _ENGINE_SCHEMA, mismatched, rows_b[:1], True,
            tmp / "route-mismatch.xlsx", mode="values")
        assert mismatch_result.status == "error"
        assert "outer route" in mismatch_result.message
        mixed_result = run_compare(
            _ENGINE_SCHEMA, rows_a, legacy_b, True,
            tmp / "mixed-mode.xlsx", mode="values")
        assert mixed_result.status == "error"
        assert "every key" in mixed_result.message


def test_canonical_display_is_side_order_independent():
    rows_a, rows_b = _engine_rows()
    canonical_display = "001 / ORA / R1.000"
    assert all(row[2].physical_identity.display == canonical_display
               for row in rows_a + rows_b)
    assert not (rows_a[0][2] < rows_b[0][2])
    assert not (rows_b[0][2] < rows_a[0][2])
    later_identity = cc.make_physical_identity(
        "001", "ORA", "R2.000", _engine_claims("later", 1),
        "also ignored")
    later_key = cc.physical_key("AAA source text", later_identity)
    assert rows_a[0][2] < later_key
    assert not (later_key < rows_a[0][2])
    for compare in (
            lambda: rows_a[0][2] < "legacy",
            lambda: "legacy" < rows_a[0][2]):
        try:
            compare()
        except TypeError:
            pass
        else:
            raise AssertionError("mixed PhysicalKey ordering was accepted")

    with TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        for label, side_a, side_b in (
                ("ab", rows_a, rows_b), ("ba", rows_b, rows_a)):
            output = tmp / f"{label}.xlsx"
            result = run_compare(
                _ENGINE_SCHEMA, side_a, side_b, True, output, mode="values")
            assert result.status == "ok", result.message
            wb = load_workbook(output, read_only=False, data_only=False)
            try:
                assert [wb["Comparison"].cell(row, 2).value
                        for row in (2, 3)] == [
                            canonical_display, canonical_display]
                assert wb["Spot Check"]["F7"].value == canonical_display
                assert wb["A"]["D2"].value == str(side_a[0][2])
            finally:
                wb.close()


def test_highway_sequence_env_county_and_glued_pm():
    # CMP-AUD-045, GREEN since the Highway Sequence projector batch. The
    # canonical postmile is the oracle-approved COMPLETE GLUED form
    # ("R001.000E" — zero padding, realignment prefix, AND equate suffix, per
    # the Stage-8 HSL comparison oracle's Row.identity) — unlike RD (norm_pm)
    # and ID (PP + Decimal PM), so never assume the canonical form across
    # families. County folds into the identity; prefix/PM/suffix stay
    # conserved raw claims.
    header = [
        "County", "City", "(col C)", "PM", "(col E)", "HG", "FT",
        "Distance To Next Point", "Description",
    ]
    schema = compare_env.HIGHWAY_SEQUENCE._schema(header, "A", "B")
    rows_a = [
        ["001", "LA", "CITY-A", "R", "001.000", "E", "D", "FT-A", "0.1", "DESC-A"],
        ["001", "ORA", "CITY-B", "R", "001.000", "E", "U", "FT-B", "0.2", "DESC-B"],
    ]
    rows_b = [
        ["001", "LA", "CITY-B", "R", "001.000", "E", "U", "FT-B", "0.2", "DESC-B"],
        ["001", "ORA", "CITY-A", "R", "001.000", "E", "D", "FT-A", "0.1", "DESC-A"],
    ]
    _assert_exact_physical_swap(
        schema, rows_a, rows_b, 10, "R001.000E")


def test_ramp_env_county_identity():
    # CMP-AUD-045, GREEN since the RD projector batch. The canonical postmile is
    # the owner-approved D4 form `norm_pm(PM)` — NOT the glued prefix/PM/suffix
    # this contract originally expected: the accepted RD-79 oracle keys on
    # exactly (Route, County, norm_pm(PM)) with PR/PM_SFX as separately
    # conserved claims, PR differs on zero paired rows corpus-wide, and TSN's
    # 313 print suffixes have no TSMIS counterpart — gluing either would
    # fabricate one-sided rows out of physically identical ramps.
    header = [
        "Location", "(col B)", "PM", "Date of Record", "(col E)", "HG",
        "Area 4", "(col H)", "City Code", "R/U", "Description",
    ]
    schema = compare_env.RAMP_DETAIL._schema(header, "A", "B")
    rows_a = [
        ["101", "01-DN-101", "R", "001.000", "2026-01-01", "E", "D", "Y", "", "A", "U", "DESC-A"],
        ["101", "07-LA-101", "R", "001.000", "2026-01-01", "E", "U", "N", "", "B", "R", "DESC-B"],
    ]
    rows_b = [
        ["101", "01-DN-101", "R", "001.000", "2026-01-01", "E", "U", "N", "", "B", "R", "DESC-B"],
        ["101", "07-LA-101", "R", "001.000", "2026-01-01", "E", "D", "Y", "", "A", "U", "DESC-A"],
    ]
    _assert_exact_physical_swap(
        schema, rows_a, rows_b, 10, "1.000")
    # The conserved raw claims carry the prefix/suffix facts the key excludes.
    keys = keys_for(rows_a, True, schema.key_field, schema.key_normalizer)
    claims = dict((c.name, c.value)
                  for c in keys[0][1].physical_identity.raw_claims)
    assert claims == {"route": "101", "location": "01-DN-101",
                      "postmile_prefix": "R", "postmile": "001.000",
                      "postmile_suffix": "E"}, claims


def test_intersection_env_county_identity():
    # CMP-AUD-045, GREEN since the ID projector batch. The canonical postmile is
    # the accepted ID-79 form `complete PP + Decimal-canonical PM` ("R1" for
    # PP=R, PM=001.000) — NOT the glued padded-PM+suffix this contract
    # originally expected: the accepted oracle's tuple is (base Route, County,
    # complete PP, numeric Post Mile), with the route/PM SUFFIX a conserved
    # claim only.
    header = ["P", "Post Mile", "S", "Location", "City Code", "R/U", "Description"]
    schema = compare_env.INTERSECTION_DETAIL._schema(header, "A", "B")
    rows_a = [
        ["001", "R", "001.000", "E", "01 DN 001", "A", "U", "DESC-A"],
        ["001", "R", "001.000", "E", "07 LA 001", "B", "R", "DESC-B"],
    ]
    rows_b = [
        ["001", "R", "001.000", "E", "01 DN 001", "B", "R", "DESC-B"],
        ["001", "R", "001.000", "E", "07 LA 001", "A", "U", "DESC-A"],
    ]
    _assert_exact_physical_swap(
        schema, rows_a, rows_b, 6, "R1")
    # PP is PART of identity for ID (six real within-county groups): the same
    # county+numeric-PM under a DIFFERENT complete PP is a different location.
    keys = keys_for(rows_a, True, schema.key_field, schema.key_normalizer)
    other_pp = [list(rows_a[0]), list(rows_a[0])]
    other_pp[1][1] = "M"                          # same county/PM, PP R vs M
    keys_pp = keys_for(other_pp, True, schema.key_field, schema.key_normalizer)
    assert keys_pp[0][1] != keys_pp[1][1], "distinct complete PPs must not collide"
    claims = dict((c.name, c.value)
                  for c in keys[0][1].physical_identity.raw_claims)
    assert claims == {"route": "001", "route_suffix": "",
                      "location": "01 DN 001", "postmile_prefix": "R",
                      "postmile": "001.000", "postmile_suffix": "E"}, claims


def _claims(*items):
    return tuple(cc.RawIdentityClaim(name, value) for name, value in items)


def _assert_identity(row, *, key_field, route, county, postmile, raw_claims):
    key_index = 1 + key_field
    assert key_index < len(row), ("projector omitted engine key field", row)
    assert type(row[key_index]) is cc.PhysicalKey, (
        "engine-consumed key cell is not PhysicalKey", key_index, row)
    identity = cc.physical_identity_from_key(row, 1, key_field)
    assert identity is row[key_index].physical_identity
    components = dict(identity.canonical_components)
    assert components == {
        "route": route,
        "county": county,
        "postmile": postmile,
    }, components
    assert identity.raw_claims == raw_claims, (
        "projector raw claims are not exact/lossless",
        identity.raw_claims,
        raw_claims,
    )


_RD_CLAIMS = _claims(
    ("route", "101"), ("location", "01-DN-101"),
    ("postmile_prefix", "R"), ("postmile", "001.000"),
    ("postmile_suffix", "E"))


def test_ramp_direct_projectors_retain_physical_identity():
    """CMP-AUD-045, GREEN since the RD projector batch: every direct Ramp Detail
    path bakes the D4 PhysicalKey (canonical postmile = `norm_pm(PM)`, per the
    accepted RD-79 oracle — prefix/suffix are conserved claims, never key
    components; see test_ramp_env_county_identity) and now carries District as
    a compared field (CMP-AUD-185) at header index 2."""
    ramp_tsmis = rd._tsmis_row([
        "101", "01-DN-101", "R", "001.000", "2026-01-01", "E", "D", "Y", "CITY", "U", "DESC", "",
    ])
    _assert_identity(
        ramp_tsmis, key_field=rd.KEY_FIELD,
        route="101", county="DN", postmile="1.000",
        raw_claims=_RD_CLAIMS)
    assert rd.SHARED_HEADER[2] == "District" and ramp_tsmis[3] == "01", ramp_tsmis

    ramp_raw_header = {
        "LOCATION": 0, "PR": 1, "PM": 2, "PM_SFX": 3,
        "DATE_OF_RECORD": 4, "HG": 5, "AREA_4": 6, "CITY_CODE": 7,
        "POP": 8, "DESCRIPTION": 9,
    }
    ramp_tsn = rd._tsn_raw_row(
        ["01-DN-101", "R", "001.000", "E", "2026-01-01", "D", "Y", "CITY", "U", "12/DESC"],
        ramp_raw_header,
    )
    _assert_identity(
        ramp_tsn, key_field=rd.KEY_FIELD,
        route="101", county="DN", postmile="1.000",
        raw_claims=_RD_CLAIMS)
    # CMP-AUD-135: TSN's own text is authoritative — the leading "12/" survives.
    assert ramp_tsn[1 + rd.SHARED_HEADER.index("Description")] == "12/DESC", ramp_tsn

    ramp_pdf = rdpdf._pdf_row([
        "101", "01-DN-101", "R", "001.000", "2026-01-01", "E",
        "D", "Y", "CITY", "U", "DESC", "", "N", "TYPE",
    ])
    _assert_identity(
        ramp_pdf, key_field=rd.KEY_FIELD,
        route="101", county="DN", postmile="1.000",
        raw_claims=_RD_CLAIMS)
    assert ramp_pdf[3] == "01", ramp_pdf


def test_id_direct_projector_retains_physical_identity():
    """CMP-AUD-045, GREEN since the ID projector batch: every direct
    Intersection Detail path (the PDF flavor reuses this same loader) bakes the
    accepted ID-79 PhysicalKey — canonical postmile = complete PP +
    Decimal-canonical PM ("R1" for PP=R, PM=001.000); the route/PM suffixes and
    Location stay conserved claims."""
    intersection = [None] * 36
    intersection[0] = "001"
    intersection[1] = "R"
    intersection[2] = "001.000"
    intersection[3] = "E"
    intersection[4] = "01 DN 001"
    id_row = idt._tsmis_row(intersection)
    _assert_identity(
        id_row, key_field=idt.KEY_FIELD,
        route="001", county="DN", postmile="R1",
        raw_claims=_claims(
            ("route", "001"), ("route_suffix", ""),
            ("location", "01 DN 001"),
            ("postmile_prefix", "R"), ("postmile", "001.000"),
            ("postmile_suffix", "E")))

    tsn_header = {"LOCATION": 0, "PP": 1, "POST_MILE": 2, "HG": 3,
                  "DESCRIPTION": 4}
    tsn_row = idt._tsn_row(["01 DN 001", "R", "001.000", "D", "DESC"],
                           tsn_header)
    _assert_identity(
        tsn_row, key_field=idt.KEY_FIELD,
        route="001", county="DN", postmile="R1",
        raw_claims=_claims(
            ("route", "001"), ("route_suffix", ""),
            ("location", "01 DN 001"),
            ("postmile_prefix", "R"), ("postmile", "001.000")))
    # Decimal canonical: '005.870' -> '5.87'; zero -> '0'.
    assert idt._decimal_pm("005.870") == "5.87"
    assert idt._decimal_pm("0.000") == "0"


def test_hsl_direct_projector_retains_physical_identity():
    """CMP-AUD-045/199, GREEN since the Highway Sequence projector batch: the
    vs-TSN projector bakes the complete-glued-postmile PhysicalKey (suffix IN
    the canonical, HSL's own convention), the TSN loader builds the same
    identity — with the 46 blank-County raw annotations keyed under the
    explicit "(county not printed)" marker (CMP-AUD-158), never dropped — and
    the PDF-vs-Excel same-source projector keys WITHOUT the equate suffix
    (CMP-AUD-199: the two renders seat "E" on different rows of one equate
    pair, so it is a compared "PM Suffix" cell there, not identity)."""
    hsl_row = hsl._tsmis_row([
        "001", "ORA.", "CITY", "R", "001.000", "E", "D", "H", "0.1", "DESC",
    ])
    _assert_identity(
        hsl_row, key_field=hsl.KEY_FIELD,
        route="001", county="ORA", postmile="R001.000E",
        raw_claims=_claims(
            ("route", "001"), ("county", "ORA."),
            ("postmile_prefix", "R"), ("postmile", "001.000"),
            ("postmile_suffix", "E")))

    # The same-source (PDF vs Excel) projector: suffix excluded from identity,
    # exposed as the compared "PM Suffix" column; descriptions verbatim.
    ss_row = hslpdf._tsmis_row_same_source([
        "001", "ORA.", "CITY", "R", "001.000", "E", "D", "H", "0.1", "1/DESC",
    ])
    _assert_identity(
        ss_row, key_field=hslpdf.SS_KEY_FIELD,
        route="001", county="ORA", postmile="R001.000",
        raw_claims=_claims(
            ("route", "001"), ("county", "ORA."),
            ("postmile_prefix", "R"), ("postmile", "001.000"),
            ("postmile_suffix", "E")))
    assert ss_row[1 + hslpdf.SS_HEADER.index("PM Suffix")] == "E", ss_row
    assert ss_row[1 + hslpdf.SS_HEADER.index("Description")] == "1/DESC", ss_row
    assert hslpdf._SS_SCHEMA.context_fields == ()

    # The TSN loader's explicit unknown-county disclosure (CMP-AUD-158): a raw
    # annotation printed before county context keys under the reserved marker —
    # it can never pair with a real county, and it is never dropped.
    key = hsl._physical_pm_key(
        "001", None, "R000.000", (("route", "001"), ("county", ""),
                                  ("postmile", "R000.000")), "probe")
    components = dict(key.physical_identity.canonical_components)
    assert components["county"] == "(county not printed)", components


TESTS = (
    ("engine preserves exact per-side PhysicalKey claims",
     test_engine_preserves_each_side_raw_claim),
    ("LoadedSide tagged roundtrip and raw-claim scalar domain",
     test_loaded_side_roundtrip_and_claim_domain),
    ("canonical route authority and uniform typed mode",
     test_route_authority_and_uniform_typed_mode),
    ("canonical display and ordering are side-independent",
     test_canonical_display_is_side_order_independent),
    # Promoted from KNOWN_RED by the Ramp Detail projector batch (CMP-AUD-045/
    # 135/185): D4 county-aware keys on every RD path, District compared, raw
    # TSN Descriptions preserved.
    ("Ramp env uses county identity", test_ramp_env_county_identity),
    ("Ramp direct/PDF projectors retain structured identity",
     test_ramp_direct_projectors_retain_physical_identity),
    # Promoted by the Intersection Detail projector batch (CMP-AUD-045-ID):
    # the accepted 4-part tuple with complete PP inside the canonical postmile.
    ("Intersection env uses county identity",
     test_intersection_env_county_identity),
    ("Intersection direct/PDF projector retains structured identity",
     test_id_direct_projector_retains_physical_identity),
    # Promoted by the Highway Sequence projector batch (CMP-AUD-045/158/199):
    # complete glued-postmile identity on the direct/env/TSN paths, the
    # blank-county/blank-postmile reserved markers, and the same-source
    # (PDF vs Excel) suffix-excluded identity with "PM Suffix" compared.
    ("Highway Sequence env uses county + glued PM",
     test_highway_sequence_env_county_and_glued_pm),
    ("HSL direct/same-source projectors retain structured identity",
     test_hsl_direct_projector_retains_physical_identity),
)

# CMP-AUD-045 red fixture, now fully promoted: all four report families
# (RD, ID, HSL ×2 contracts) graduated into TESTS above. The mechanism stays
# for the next finding that needs a live red fixture: each entry is
# (label, test, exact-failure-signature) — it must FAIL with that signature
# until fixed, then be promoted into TESTS.
KNOWN_RED = ()


def main():
    failures = []
    for label, test in TESTS:
        try:
            test()
        except Exception as exc:
            failures.append((label, exc))
            print(f"FAIL {label}: {type(exc).__name__}: {exc}")
        else:
            print(f"OK   {label}")

    for label, test, signature in KNOWN_RED:
        try:
            test()
        except AssertionError as exc:
            if signature in str(exc):
                print(f"RED  {label}  (known: CMP-AUD-045)")
            else:
                failures.append((label, exc))
                print(f"FAIL {label}: CMP-AUD-045 red fixture DRIFTED — expected "
                      f"{signature!r}, got: {exc}")
        except Exception as exc:
            failures.append((label, exc))
            print(f"FAIL {label}: CMP-AUD-045 red fixture raised an unexpected "
                  f"{type(exc).__name__}: {exc}")
        else:
            failures.append((label, None))
            print(f"FAIL {label}: CMP-AUD-045 looks FIXED here — move this contract "
                  f"from KNOWN_RED into TESTS and re-run the owning family gate")

    if failures:
        raise SystemExit(f"{len(failures)} physical-identity contract(s) failed")
    print(f"OK  PHASE4-L0A-PHYSICAL-IDENTITY "
          f"({len(TESTS)} green, {len(KNOWN_RED)} known-red under CMP-AUD-045)")


if __name__ == "__main__":
    main()
