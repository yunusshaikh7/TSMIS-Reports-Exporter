"""Canonical TSN library — the ONE fixed home the whole app references for each
report's TSN source data.

The six TSN reports essentially never change, so rather than scatter them across
per-run drop folders (the legacy ``input/tsn_highway_log/`` and the matrix's
``<dest>/_tsn_input/<subdir>/``), each report's TSN data lives under
``<DATA_ROOT>/tsn_library/<report>/``:

    raw/           the RAW TSN file(s) exactly as exported — district PDFs, a
                   statewide PDF, or a statewide XLSX (the format is the report's
                   own, see tsn-parsers.md).
    consolidated/  the GENERATED consolidated / normalized Excel, built once from
                   the raw and reused for every comparison. A couple of TSN
                   reports are PDFs, so this Excel is what the comparison engine
                   actually reads.

Every consolidator / comparator / matrix DEFAULTS to this library — the per-row
``tsn_subdir`` IS the report key here. An explicit user file-pick stays an
override (``resolve(report, selected_file=...)``). The legacy Highway Log
locations are honored as read-only fallbacks so existing installs keep working
until imported.

Console-free (no print / input / sys.exit): functions return dicts /
ConsolidateResult or raise. The report's own builder/consolidator imports are LAZY
(resolved inside build_consolidated). As of P4 this module imports `report_catalog`
to derive its descriptors, which (like the registry) transitively pulls openpyxl /
pdfplumber — so importing tsn_library is console-free but not dependency-light.
"""
import hashlib
import importlib
import json
import logging
import math
import os
import shutil
import stat
import tempfile
from dataclasses import dataclass
from pathlib import Path

import consolidation_meta
import outcome
import paths
import report_catalog as _catalog
import tsn_district_contract as _raw_contract
from events import ConsolidateResult, Events

log = logging.getLogger("tsmis.tsn_library")


@dataclass(frozen=True)
class TsnReport:
    """One TSN report's library descriptor — the single place that knows its raw
    format and how to build its consolidated/normalized Excel."""
    subdir: str            # report key == the matrix's per-row tsn_subdir token
    label: str             # human label for the Settings status panel
    raw_glob: str          # "*.pdf" | "*.xlsx" — what counts as a raw file
    raw_kind: str          # "district_pdfs" | "statewide_pdf" | "statewide_xlsx"
    consolidated_name: str # filename written into consolidated/
    builder: str           # "module:function"; func(raw_dir, out_path, events, confirm_overwrite)
    normalization_version: int = 1   # bumped in the CATALOG when the normalizer changes (D2)
    evidence_pdfs: bool = False      # also keep a pdf/ drop folder (evidence images)


# Registry — DERIVED from report_catalog (the report-metadata SoT, P4), so each
# report's TSN descriptor lives in exactly one place. Highway Log shipped first in
# v0.17.0; the format facts come from the raw 6.19 ground-truth set (see
# docs/tsn-parsers.md). Keyed by subdir, in catalog (registration) order.
_REPORTS = {
    e.subdir: TsnReport(
        subdir=e.subdir,
        label=e.label,
        raw_glob=e.raw_glob,
        raw_kind=e.raw_kind,
        consolidated_name=e.consolidated_name,
        builder=e.builder,
        normalization_version=e.normalization_version,
        evidence_pdfs=e.evidence_pdfs,
    )
    for e in _catalog.tsn_entries()
}


# --------------------------------------------------------------------------- #
# Registry accessors
# --------------------------------------------------------------------------- #
def reports():
    """Every registered TSN report descriptor, in registration order."""
    return list(_REPORTS.values())


def is_registered(report):
    return report in _REPORTS


def get(report):
    """The descriptor for `report`, or raise ValueError for an unknown key."""
    try:
        return _REPORTS[report]
    except KeyError:
        raise ValueError(f"unknown TSN report: {report!r}")


# --------------------------------------------------------------------------- #
# Filesystem layout
# --------------------------------------------------------------------------- #
def raw_dir(report):
    return paths.tsn_library_raw_dir(report)


def pdf_dir(report):
    """The report's OPTIONAL TSN district-print drop folder (evidence images);
    meaningful only for entries with `evidence_pdfs=True`."""
    return paths.tsn_library_pdf_dir(report)


def consolidated_path(report):
    return paths.tsn_library_consolidated_path(report, get(report).consolidated_name)


def _safe_mtime(p):
    try:
        return Path(p).stat().st_mtime
    except OSError:               # silent-ok: a pure mtime probe; None = unknown
        return None


def _raw_probe(report):
    """Return ``(ordinary matching members, probe_error)`` for one raw folder.

    An unreadable folder is materially different from an empty folder once a
    consolidated workbook exists. Preserve that distinction through
    ``status``/``ensure_current`` so comparison consumers receive an actionable
    fail-closed result instead of reading an uncertified workbook.
    """
    try:
        candidates = raw_dir(report).glob(get(report).raw_glob)
        members = sorted(
            (p for p in candidates
             if not p.name.startswith("~$") and stat.S_ISREG(p.stat().st_mode)),
            key=lambda p: p.name.casefold())
    except OSError as e:
        log.warning("TSN raw folder unreadable for %s (%s: %s)",
                    report, type(e).__name__, e)
        return [], f"{type(e).__name__}: {e}"
    return members, None


def _raw_files(report):
    """Sorted raw files, preserving the historical list-only helper surface."""
    return _raw_probe(report)[0]


def _raw_manifest(report, raws=None):
    """Strict canonical content identity for one report's ordinary raw members."""
    members = list(_raw_files(report) if raws is None else raws)
    return _raw_contract.canonical_raw_manifest(members, raw_dir(report))


_NORMALIZED_WORKBOOK_IDENTITY_VERSION = 1
_TSN_ARTIFACT_IDENTITY_VERSION = 1
_TSN_SIDECAR_MAX_BYTES = 1024 * 1024
_NORMALIZED_IDENTITY_KEYS = {
    "version", "algorithm", "byte_length", "sha256",
}


def _sha256_text(value):
    return (isinstance(value, str) and len(value) == 64
            and all(char in "0123456789abcdef" for char in value))


def _finite_number(value):
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    try:
        return math.isfinite(float(value))
    except (OverflowError, TypeError, ValueError):  # silent-ok: pure shape predicate, not an error path
        return False


def _bound_stat_signature(value):
    """Mutation-sensitive descriptor/path signature for a durable file read."""
    return (
        int(getattr(value, "st_dev", 0)),
        int(getattr(value, "st_ino", 0)),
        stat.S_IFMT(value.st_mode),
        int(value.st_size),
        int(getattr(value, "st_mtime_ns", value.st_mtime * 1e9)),
        int(getattr(value, "st_ctime_ns", value.st_ctime * 1e9)),
    )


def _open_bound_file(path):
    """Open an ordinary file and bind its descriptor to the current pathname."""
    path = Path(path)
    try:
        before = path.lstat()
    except OSError as exc:
        raise ValueError(
            f"file could not be inspected ({type(exc).__name__}: {exc})") from exc
    if not stat.S_ISREG(before.st_mode):
        raise ValueError("path is not an ordinary file")
    flags = (os.O_RDONLY | getattr(os, "O_BINARY", 0)
             | getattr(os, "O_NOINHERIT", 0) | getattr(os, "O_NOFOLLOW", 0))
    try:
        descriptor = os.open(path, flags)
    except (OSError, ValueError) as exc:
        raise ValueError(
            f"file could not be opened ({type(exc).__name__}: {exc})") from exc
    try:
        opened = os.fstat(descriptor)
        if (not stat.S_ISREG(opened.st_mode)
                or _bound_stat_signature(opened) != _bound_stat_signature(before)):
            raise ValueError("file identity changed while opening")
    except Exception:
        try:
            os.close(descriptor)
        except OSError:  # silent-ok: cleanup close; the bare re-raise below carries the real cause
            pass
        raise
    return path, descriptor, opened


def _finish_bound_file(path, descriptor, opened):
    """Close and prove the read descriptor still names the live stable entry."""
    close_error = None
    try:
        after_descriptor = os.fstat(descriptor)
    except OSError as exc:
        raise ValueError("file identity could not be verified after reading") from exc
    finally:
        try:
            os.close(descriptor)
        except OSError as exc:  # silent-ok: not swallowed -- re-raised as ValueError just below
            close_error = exc
    if close_error is not None:
        raise ValueError("file descriptor could not be closed after reading") from close_error
    try:
        after_path = path.lstat()
    except OSError as exc:
        raise ValueError("file path changed after reading") from exc
    expected = _bound_stat_signature(opened)
    if (_bound_stat_signature(after_descriptor) != expected
            or _bound_stat_signature(after_path) != expected):
        raise ValueError("file changed while it was being read")


def normalized_workbook_identity(path):
    """Return a stable content identity for one normalized workbook.

    The digest is read through an ordinary-file descriptor bound to the current
    pathname. Descriptor and pathname identity/size/timestamps must remain exact
    through the read, so a concurrent replacement never receives a generation
    claim for whichever bytes happened to be observed.
    """
    path, descriptor, opened = _open_bound_file(path)
    digest = hashlib.sha256()
    count = 0
    try:
        while True:
            block = os.read(descriptor, 1024 * 1024)
            if not block:
                break
            digest.update(block)
            count += len(block)
    except OSError as exc:
        try:
            os.close(descriptor)
        except OSError:  # silent-ok: cleanup close; the ValueError below carries the real cause
            pass
        raise ValueError(
            f"normalized workbook could not be hashed ({type(exc).__name__}: {exc})") from exc
    _finish_bound_file(path, descriptor, opened)
    if count != opened.st_size:
        raise ValueError("normalized workbook size changed while hashing")
    return {
        "version": _NORMALIZED_WORKBOOK_IDENTITY_VERSION,
        "algorithm": "sha256",
        "byte_length": count,
        "sha256": digest.hexdigest(),
    }


def validate_normalized_workbook_identity(value):
    """Return the strict normalized-workbook identity or raise ``ValueError``."""
    if not isinstance(value, dict) or set(value) != _NORMALIZED_IDENTITY_KEYS:
        raise ValueError("normalized workbook identity shape is invalid")
    size = value.get("byte_length")
    if (not isinstance(value.get("version"), int)  # CMP-AUD-035: reject 1.0 aliasing 1
            or isinstance(value.get("version"), bool)
            or value.get("version") != _NORMALIZED_WORKBOOK_IDENTITY_VERSION
            or value.get("algorithm") != "sha256"
            or not isinstance(size, int) or isinstance(size, bool) or size < 0
            or not _sha256_text(value.get("sha256"))):
        raise ValueError("normalized workbook identity is invalid")
    return {
        "version": _NORMALIZED_WORKBOOK_IDENTITY_VERSION,
        "algorithm": "sha256",
        "byte_length": size,
        "sha256": value["sha256"],
    }


def canonical_normalized_identity_token(report, raw_manifest,
                                        workbook_identity):
    """Deterministically bind one TSN normalized artifact generation.

    The returned token covers the canonical dataset key, catalog normalizer
    version, complete raw-member manifest, and exact normalized workbook bytes.
    It is deliberately independent of mtimes and absolute installation paths.
    """
    spec = get(report)
    raw_manifest = _raw_contract.validate_raw_manifest(raw_manifest)
    workbook_identity = validate_normalized_workbook_identity(workbook_identity)
    payload = {
        "version": _TSN_ARTIFACT_IDENTITY_VERSION,
        "dataset": spec.subdir,
        "normalization_version": spec.normalization_version,
        "raw_manifest": raw_manifest,
        "normalized_workbook": workbook_identity,
    }
    canonical = json.dumps(
        payload, ensure_ascii=False, sort_keys=True,
        separators=(",", ":"), allow_nan=False).encode("utf-8")
    return (f"tsn-normalized-v{_TSN_ARTIFACT_IDENTITY_VERSION}:"
            f"{hashlib.sha256(canonical).hexdigest()}")


def _bound_sidecar_payload(path):
    """Read one small TSN sidecar as strict JSON through a bound descriptor."""
    path, descriptor, opened = _open_bound_file(path)
    if opened.st_size > _TSN_SIDECAR_MAX_BYTES:
        os.close(descriptor)
        raise ValueError("TSN sidecar exceeds its size limit")
    raw = bytearray()
    try:
        while len(raw) <= _TSN_SIDECAR_MAX_BYTES:
            block = os.read(
                descriptor,
                min(1024 * 1024, _TSN_SIDECAR_MAX_BYTES + 1 - len(raw)))
            if not block:
                break
            raw.extend(block)
    except OSError as exc:
        try:
            os.close(descriptor)
        except OSError:  # silent-ok: cleanup close; the ValueError below carries the real cause
            pass
        raise ValueError("TSN sidecar could not be read completely") from exc
    _finish_bound_file(path, descriptor, opened)
    if len(raw) > _TSN_SIDECAR_MAX_BYTES or len(raw) != opened.st_size:
        raise ValueError("TSN sidecar changed size while being read")

    def unique_object(pairs):
        result = {}
        for key, value in pairs:
            if key in result:
                raise ValueError(f"duplicate JSON key {key!r}")
            result[key] = value
        return result

    def reject_constant(value):
        raise ValueError(f"non-finite JSON number {value!r}")

    try:
        value = json.loads(
            bytes(raw).decode("utf-8"), object_pairs_hook=unique_object,
            parse_constant=reject_constant)
    except (UnicodeError, ValueError, RecursionError) as exc:
        raise ValueError("TSN sidecar is not strict UTF-8 JSON") from exc
    if not isinstance(value, dict):
        raise ValueError("TSN sidecar must be a JSON object")
    return value


# --------------------------------------------------------------------------- #
# On-disk skeleton — make a fresh/empty library self-documenting so the user
# can SEE where each report's TSN files go before importing the first one. The
# hint/README files are plain .txt, so the *.pdf/*.xlsx readers never see them.
# --------------------------------------------------------------------------- #
_README_NAME = "_README - where TSN files go.txt"
_RAW_HINT_NAME = "_PUT TSN FILES HERE.txt"

_RAW_KIND_DESC = {
    "district_pdfs": "the per-district PDF files (one PDF per Caltrans district)",
    "statewide_pdf": "the single statewide PDF",
    "statewide_xlsx": "the single statewide Excel workbook",
}


def _raw_kind_desc(spec):
    return _RAW_KIND_DESC.get(spec.raw_kind, f"the raw TSN file(s) ({spec.raw_glob})")


def _raw_hint_text(spec):
    return (
        f"{spec.label} - TSN source files\n"
        f"{'=' * 48}\n\n"
        f"Put {_raw_kind_desc(spec)} here.\n"
        f"Expected file type: {spec.raw_glob}\n\n"
        f"Then, in the app: Settings -> TSN reports -> Rebuild ({spec.label}),\n"
        "which builds the workbook the comparison actually reads. (The Import\n"
        "button there can also copy the file(s) in here for you.)\n\n"
        "This note is ignored by the app and is safe to delete.\n"
    )


def _pdf_hint_text(spec):
    return (
        f"{spec.label} - district PDFs (optional, for Evidence images)\n"
        f"{'=' * 48}\n\n"
        "Drop the TSN DISTRICT print PDFs here (any filenames - the app reads\n"
        "each file's own DIST-CNTY-ROUTE header to know its district).\n\n"
        "They enable the 'Evidence images' option on the comparison pages:\n"
        "sampled differences rendered as highlighted snippets from both the\n"
        "TSMIS and TSN printed pages. The comparison itself never reads these\n"
        "(the raw/ workbook is its source); no rebuild is needed after adding\n"
        "or replacing them.\n\n"
        "This note is ignored by the app and is safe to delete.\n"
    )


def _readme_text():
    lines = [
        "TSMIS Exporter - TSN report library",
        "=" * 48,
        "",
        "This folder is where the app looks for each report's TSN source data",
        "(the 'ground truth' your TSMIS exports are compared against).",
        "",
        "Each report has its own folder. Drop that report's TSN file(s) into its",
        "'raw' subfolder, then rebuild it in  Settings -> TSN reports.",
        "",
        "Report folder  ->  what to drop in its raw/ subfolder:",
        "",
    ]
    for spec in reports():
        lines.append(f"  {spec.subdir}/raw/   {spec.label}")
        lines.append(f"      {_raw_kind_desc(spec)}  ({spec.raw_glob})")
    lines += [
        "",
        "The 'consolidated' subfolder is generated by the app — leave it alone.",
    ]
    for spec in reports():
        if spec.evidence_pdfs:
            lines += [
                "",
                f"OPTIONAL: {spec.subdir}/pdf/ — drop the {spec.label} DISTRICT",
                "print PDFs there (any filenames) to enable the 'Evidence images'",
                "option on the comparison pages: sampled differences rendered as",
                "highlighted snippets from both the TSMIS and TSN printed pages.",
            ]
    lines += [
        "",
        "This note is ignored by the app and is safe to delete.",
    ]
    return "\n".join(lines) + "\n"


def ensure_layout():
    """Create the on-disk skeleton so an empty library is self-documenting: the
    root, each report's raw/ folder (+ the pdf/ evidence drop folder where the
    catalog says one exists), a per-folder hint file naming what goes there, and
    a root README mapping every report to its folder. Idempotent and
    best-effort — never overwrites raw data, only seeds the hint into a raw/
    folder that has no real files yet (so it won't fight a user who deleted it
    after adding data), and swallows OSError. The README refreshes whenever its
    GENERATED text changed (v0.21.1 — an updated install must learn about new
    folders; it stays user-deletable between launches only until the next run,
    same as before). Returns the library root Path."""
    root = paths.TSN_LIBRARY_ROOT
    try:
        root.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        log.info("TSN library root not creatable (%s: %s)", type(e).__name__, e)
        return root
    for spec in reports():
        rd = raw_dir(spec.subdir)
        try:
            rd.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            log.info("TSN raw dir not creatable for %s (%s: %s)",
                     spec.subdir, type(e).__name__, e)
            continue
        if not _raw_files(spec.subdir):                  # only hint an empty raw/
            hint = rd / _RAW_HINT_NAME
            if not hint.exists():
                try:
                    hint.write_text(_raw_hint_text(spec), encoding="utf-8")
                except OSError:   # silent-ok: the hint file is cosmetic guidance
                    pass
        if spec.evidence_pdfs:
            pd = pdf_dir(spec.subdir)
            try:
                pd.mkdir(parents=True, exist_ok=True)
            except OSError as e:
                log.info("TSN pdf dir not creatable for %s (%s: %s)",
                         spec.subdir, type(e).__name__, e)
                continue
            if not any(pd.glob("*.pdf")):                # only hint an empty pdf/
                hint = pd / "_PUT TSN DISTRICT PDFS HERE.txt"
                if not hint.exists():
                    try:
                        hint.write_text(_pdf_hint_text(spec), encoding="utf-8")
                    except OSError:   # silent-ok: cosmetic guidance
                        pass
    readme = root / _README_NAME
    try:
        current = _readme_text()
        stale = (not readme.exists()
                 or readme.read_text(encoding="utf-8") != current)
        if stale:
            readme.write_text(current, encoding="utf-8")
    except OSError:               # silent-ok: the README is cosmetic guidance
        pass
    return root


# --------------------------------------------------------------------------- #
# Legacy fallbacks (Highway Log only) — honored read-only until the user imports
# the raw into the library, so existing installs never break. Computed lazily from
# paths.* at call time (so a test that redirects OUTPUT_ROOT/INPUT_ROOT is honored).
# --------------------------------------------------------------------------- #
def _legacy_consolidated(report):
    if report == "highway_log":
        return paths.OUTPUT_ROOT / "tsn_highway_log_consolidated.xlsx"
    return None


def _legacy_raw_dir(report):
    if report == "highway_log":
        return paths.INPUT_ROOT / "tsn_highway_log"
    return None


def _resolve_legacy_global(report):
    """The pre-library global locations for `report` (consolidated workbook, then
    raw drop folder). Returns the tsn_source-shaped dict, or {kind:none}."""
    cons = _legacy_consolidated(report)
    if cons and cons.is_file():
        return {"kind": "consolidated", "path": str(cons),
                "mtime": _safe_mtime(cons), "legacy": True}
    raw = _legacy_raw_dir(report)
    if raw:
        n = sum(1 for p in raw.glob("*.pdf")) if raw.is_dir() else 0
        if n:
            return {"kind": "pdfs", "pdf_count": n, "legacy": True}
    return {"kind": "none"}


def _resolve_dest_drop(legacy_dest, report):
    """The matrix's dest-scoped drop <dest>/_tsn_input/<report>/ — a consolidated
    .xlsx (newest) wins, else district PDFs -> 'pdfs', else none. Mirrors
    matrix.tsn_source so an existing _tsn_input keeps resolving during cutover."""
    root = Path(legacy_dest) / "_tsn_input" / report
    xlsx, pdfs = [], 0
    try:
        for e in root.iterdir():
            if not e.is_file():
                continue
            sfx = e.suffix.lower()
            if sfx == ".xlsx" and not e.name.startswith("~$"):
                xlsx.append(e)
            elif sfx == ".pdf":
                pdfs += 1
    except OSError as e:
        log.info("TSN drop folder unreadable (%s: %s)", type(e).__name__, e)
    if xlsx:
        newest = max(xlsx, key=lambda q: _safe_mtime(q) or 0)
        return {"kind": "consolidated", "path": str(newest),
                "mtime": _safe_mtime(newest), "legacy": True}
    if pdfs:
        return {"kind": "pdfs", "pdf_count": pdfs, "legacy": True}
    return {"kind": "none"}


# --------------------------------------------------------------------------- #
# Status (drives the Settings ▸ TSN reports panel)
# --------------------------------------------------------------------------- #
def status(report):
    """{report, label, raw_present, raw_count, raw_newest_mtime,
        consolidated_present, consolidated_mtime, current}.

    `current` is True iff a consolidated workbook exists, its strict sidecar binds
    the current catalog normalizer + exact raw manifest + exact normalized workbook
    bytes, and its mtime is at least as new as every raw member. A registered report
    with no raw in the library reports raw_present=False (the user must import it)."""
    spec = get(report)
    raws, raw_probe_error = _raw_probe(report)
    raw_mtimes = [m for m in (_safe_mtime(p) for p in raws) if m is not None]
    raw_newest = max(raw_mtimes) if raw_mtimes else None
    cons = consolidated_path(report)
    cons_exists = cons.is_file()
    cons_mtime = _safe_mtime(cons) if cons_exists else None
    # D2: the library stores ALREADY-NORMALIZED values, so "current" also means
    # "built by the CURRENT normalizer". An absent/mismatched stamp (every
    # pre-stamp library, or any library built before a normalizer fix) reads
    # STALE — fail-safe; ensure_current()/build_consolidated then rebuild from
    # the retained raw instead of silently comparing stale values (the
    # v0.17.6 / v0.18.3 "looks unfixed" trap).
    # TSN uses one strict, identity-bound sidecar snapshot. Generic consolidation
    # metadata remains backward compatible, but a normalized TSN artifact is a
    # ground-truth comparison input and therefore fails closed on every missing,
    # malformed, duplicate-key, or legacy provenance field.
    certificate = None
    certificate_error = None
    if cons_exists:
        try:
            certificate = _bound_sidecar_payload(consolidation_meta.meta_path(cons))
        except ValueError as e:
            certificate_error = str(e)
        except OSError as e:
            certificate_error = f"{type(e).__name__}: {e}"
    stored_built_at = (certificate.get("built_at_mtime")
                       if isinstance(certificate, dict) else None)
    base_fields_valid = bool(
        isinstance(certificate, dict)
        and isinstance(certificate.get("schema_version"), int)  # CMP-AUD-035
        and not isinstance(certificate.get("schema_version"), bool)
        and certificate.get("schema_version") == consolidation_meta.SCHEMA_VERSION
        and certificate.get("record_type") is None
        and isinstance(certificate.get("completion"), str)
        and certificate.get("completion") in outcome.COMPLETIONS
        and isinstance(certificate.get("skipped_inputs"), int)
        and not isinstance(certificate.get("skipped_inputs"), bool)
        and certificate.get("skipped_inputs") >= 0
        and isinstance(certificate.get("failed_inputs"), int)
        and not isinstance(certificate.get("failed_inputs"), bool)
        and certificate.get("failed_inputs") >= 0
        and certificate.get("untrusted", False) is False)
    metadata_current = bool(
        cons_exists and base_fields_valid
        and _finite_number(stored_built_at) and cons_mtime is not None
        and stored_built_at == cons_mtime)
    # A provenance-valid sidecar is not necessarily a reusable producer
    # outcome.  TSN is comparison ground truth: a producer that omitted or
    # failed any input must retain its PARTIAL diagnostic, but it must never be
    # promoted to a current/green artifact merely because all identity fields
    # happen to bind its bytes.
    producer_complete = bool(
        metadata_current
        and certificate.get("completion") == outcome.COMPLETE
        and certificate.get("skipped_inputs") == 0
        and certificate.get("failed_inputs") == 0)
    stored_version = (certificate.get("tsn_normalization_version")
                      if metadata_current else None)
    norm_current = bool(
        metadata_current and isinstance(stored_version, int)
        and not isinstance(stored_version, bool)
        and stored_version == spec.normalization_version)
    # A reusable normalized artifact is current only when its present raw source
    # universe is admissible.  In particular, an already-newer workbook must not
    # bypass the exact-one gate after a second statewide candidate is dropped in.
    # District content claims are established by the builders; status can still
    # enforce their exact physical cardinality so a removed/extra member cannot
    # reuse a previously newer workbook and bypass the D01-D12 gate.
    if spec.raw_kind in ("statewide_xlsx", "statewide_pdf"):
        expected_raw_count = 1
    elif spec.raw_kind == "district_pdfs":
        expected_raw_count = 12
    else:
        expected_raw_count = None
    cardinality_admissible = bool(
        raw_probe_error is None and expected_raw_count is not None
        and len(raws) == expected_raw_count)
    current_manifest = None
    manifest_error = None
    if cardinality_admissible:
        try:
            current_manifest = _raw_manifest(report, raws)
        except (OSError, ValueError) as e:
            manifest_error = f"{type(e).__name__}: {e}"
            log.warning("TSN raw manifest unavailable for %s (%s: %s)",
                        report, type(e).__name__, e)
    raw_admissible = bool(cardinality_admissible and current_manifest is not None)
    if raw_probe_error is not None:
        raw_admission_error = f"the raw source folder could not be read ({raw_probe_error})"
    elif expected_raw_count is None:
        raw_admission_error = f"the raw source kind {spec.raw_kind!r} is unsupported"
    elif len(raws) != expected_raw_count:
        noun = "file" if expected_raw_count == 1 else "district PDF files"
        verb = "is" if expected_raw_count == 1 else "are"
        raw_admission_error = (
            f"found {len(raws)} matching raw source file(s); exactly "
            f"{expected_raw_count} {noun} {verb} required")
    elif manifest_error is not None:
        raw_admission_error = (
            f"the raw member names and bytes could not be certified ({manifest_error})")
    else:
        raw_admission_error = None
    stored_manifest = (certificate.get("tsn_raw_manifest")
                       if metadata_current else None)
    try:
        stored_manifest = _raw_contract.validate_raw_manifest(stored_manifest)
    except ValueError as e:
        # Absent is the ordinary "not built yet" state; PRESENT-but-invalid means a
        # corrupt certificate is about to force a silent rebuild -- say so.
        if stored_manifest is not None:
            log.warning("tsn: stored raw manifest for %s is unusable (%s: %s)",
                        report, type(e).__name__, e)
        stored_manifest = None
    raw_manifest_current = bool(
        raw_admissible and stored_manifest is not None
        and stored_manifest == current_manifest)

    stored_workbook_identity = (
        certificate.get("tsn_normalized_workbook_identity")
        if metadata_current else None)
    try:
        stored_workbook_identity = validate_normalized_workbook_identity(
            stored_workbook_identity)
    except ValueError as e:
        if stored_workbook_identity is not None:
            log.warning("tsn: stored normalized-workbook identity for %s is unusable "
                        "(%s: %s)", report, type(e).__name__, e)
        stored_workbook_identity = None
    current_workbook_identity = None
    workbook_identity_error = None
    if cons_exists and stored_workbook_identity is not None:
        try:
            current_workbook_identity = normalized_workbook_identity(cons)
        except ValueError as e:
            workbook_identity_error = str(e)
    normalized_workbook_current = bool(
        current_workbook_identity is not None
        and stored_workbook_identity == current_workbook_identity)

    stored_identity_token = (certificate.get("tsn_artifact_identity_token")
                             if metadata_current else None)
    expected_identity_token = None
    if (raw_manifest_current and normalized_workbook_current
            and norm_current):
        try:
            expected_identity_token = canonical_normalized_identity_token(
                report, current_manifest, current_workbook_identity)
        except ValueError as e:
            # Everything else already looks current, so a token we cannot compute is
            # abnormal: it silently demotes the library to stale.
            log.warning("tsn: canonical identity token for %s could not be computed "
                        "(%s: %s)", report, type(e).__name__, e)
            expected_identity_token = None
    identity_token_current = bool(
        isinstance(stored_identity_token, str)
        and stored_identity_token == expected_identity_token)
    candidate_current = bool(
        cons_exists and raw_admissible and cons_mtime is not None
        and raw_newest is not None and len(raw_mtimes) == len(raws)
        and cons_mtime >= raw_newest and producer_complete
        and norm_current and raw_manifest_current
        and normalized_workbook_current and identity_token_current)

    # The reads above derive one candidate claim, but they happen at different
    # instants.  A persistent replacement immediately after any one initial
    # read (including one that restores mtime) must not let status synthesize a
    # mixed-time green.  Re-read every identity-bearing component after the
    # candidate is complete and require exact agreement with that candidate.
    # This is intentionally local rather than a recursive status() call: any
    # inspection failure is a closed result, and the second observation cannot
    # accidentally derive a new generation from a different mixture.
    sidecar_revalidated = bool(candidate_current)
    raw_revalidated = bool(candidate_current)
    workbook_revalidated = bool(candidate_current)
    coherence_error = None
    if candidate_current:
        # Two complete revalidation passes catch a replacement injected after
        # either the initial observation or the first revalidation invocation.
        # No finite userspace sequence can close the instant after its final
        # filesystem read; comparison publication therefore also rechecks the
        # returned generation token at its commit boundary.
        for pass_number in (1, 2):
            pass_sidecar = False
            pass_raw = False
            pass_workbook = False
            try:
                final_certificate = _bound_sidecar_payload(
                    consolidation_meta.meta_path(cons))
                pass_sidecar = final_certificate == certificate

                final_raws, final_probe_error = _raw_probe(report)
                if final_probe_error is not None:
                    raise ValueError(
                        f"raw source folder could not be re-read ({final_probe_error})")
                initial_universe = tuple(str(Path(p).absolute()) for p in raws)
                final_universe = tuple(str(Path(p).absolute()) for p in final_raws)
                final_raw_mtimes = [_safe_mtime(p) for p in final_raws]
                if any(value is None for value in final_raw_mtimes):
                    raise ValueError("raw source mtimes could not be re-read")
                if (len(final_raws) != expected_raw_count
                        or final_universe != initial_universe):
                    raise ValueError(
                        "raw source universe changed during status inspection")
                final_manifest = _raw_manifest(report, final_raws)
                pass_raw = bool(
                    final_manifest == current_manifest
                    and final_raw_mtimes == raw_mtimes)

                final_workbook_identity = normalized_workbook_identity(cons)
                final_cons_mtime = _safe_mtime(cons)
                pass_workbook = bool(
                    final_workbook_identity == current_workbook_identity
                    and final_cons_mtime is not None
                    and final_cons_mtime == cons_mtime
                    and final_cons_mtime >= max(final_raw_mtimes))
            except (OSError, ValueError, TypeError) as e:
                coherence_error = (
                    f"revalidation pass {pass_number}: {type(e).__name__}: {e}")
                log.warning("TSN coherent status revalidation failed for %s (%s)",
                            report, coherence_error)
            sidecar_revalidated = bool(sidecar_revalidated and pass_sidecar)
            raw_revalidated = bool(raw_revalidated and pass_raw)
            workbook_revalidated = bool(
                workbook_revalidated and pass_workbook)
    coherent_snapshot_current = bool(
        candidate_current and sidecar_revalidated
        and raw_revalidated and workbook_revalidated)
    if candidate_current and not coherent_snapshot_current:
        if coherence_error is None:
            changed = []
            if not sidecar_revalidated:
                changed.append("sidecar payload")
            if not raw_revalidated:
                changed.append("raw universe/manifest")
            if not workbook_revalidated:
                changed.append("normalized workbook")
            coherence_error = "candidate changed during revalidation: " + ", ".join(changed)
        # These fields are consumer-facing current predicates too.  Keep the
        # component that actually drifted from looking independently green, and
        # always withdraw the generation token unless the whole observation was
        # coherent.
        metadata_current = bool(metadata_current and sidecar_revalidated)
        producer_complete = bool(producer_complete and sidecar_revalidated)
        norm_current = bool(norm_current and sidecar_revalidated)
        raw_manifest_current = bool(
            raw_manifest_current and sidecar_revalidated and raw_revalidated)
        normalized_workbook_current = bool(
            normalized_workbook_current
            and sidecar_revalidated and workbook_revalidated)
        identity_token_current = False
    current = coherent_snapshot_current
    return {
        "report": report,
        "label": spec.label,
        "raw_kind": spec.raw_kind,
        "raw_present": bool(raws),
        "raw_count": len(raws),
        "raw_admissible": raw_admissible,
        "raw_probe_error": raw_probe_error,
        "raw_admission_error": raw_admission_error,
        "raw_newest_mtime": raw_newest,
        "consolidated_present": cons_exists,
        "consolidated_path": str(cons),
        "consolidated_mtime": cons_mtime,
        "metadata_current": metadata_current,
        "certificate_error": certificate_error,
        "producer_complete": producer_complete,
        "normalization_current": norm_current,
        # The two numbers the "older normalizer" verdict is made of. Reported so a
        # field log distinguishes a genuinely stale stamp (3 vs 5) from a missing
        # key or a wrong type — "rebuild it" is useless advice when a rebuild has
        # just run and the report still reads stale.
        "stored_normalization_version": stored_version,
        "expected_normalization_version": spec.normalization_version,
        "raw_manifest_current": raw_manifest_current,
        "raw_manifest_sha256": (current_manifest or {}).get("sha256"),
        "normalized_workbook_current": normalized_workbook_current,
        "normalized_workbook_identity": current_workbook_identity,
        "workbook_identity_error": workbook_identity_error,
        "identity_token_current": identity_token_current,
        # Internal durable-certificate surface: unlike the consumer token below,
        # this may bind an identity-valid PARTIAL artifact so build publication
        # can verify what it wrote without making that artifact reusable.
        "certified_identity_token": (
            expected_identity_token if identity_token_current else None),
        "coherent_snapshot_current": coherent_snapshot_current,
        "coherence_error": coherence_error,
        "identity_token": expected_identity_token if current else None,
        "current": current,
    }


def all_status():
    """status() for every registered report, in registration order."""
    return [status(r.subdir) for r in reports()]


# --------------------------------------------------------------------------- #
# Import + build
# --------------------------------------------------------------------------- #
def import_raw(report, src_paths, move=False):
    """Copy (or move) user-picked raw file(s) into the report's raw/ folder — the
    only way raw enters the library. Returns the landed paths. Raises
    FileNotFoundError for a missing source; ValueError for an unknown report OR a
    file whose extension doesn't match the report's raw_glob (else it would land
    in raw/ but be invisible to the glob-based reader — a silent import-vs-ignore
    mismatch)."""
    spec = get(report)  # validate + read the expected extension
    want = spec.raw_glob.lower().lstrip("*")          # "*.pdf" -> ".pdf"
    dest = raw_dir(report)
    dest.mkdir(parents=True, exist_ok=True)
    landed = []
    for src in src_paths:
        src = Path(src)
        if not src.is_file():
            raise FileNotFoundError(f"TSN raw file not found: {src}")
        if want and src.suffix.lower() != want:
            raise ValueError(f"{src.name} is not a {want} file — {spec.label} "
                             f"expects {spec.raw_glob}.")
        target = dest / src.name
        if move:
            shutil.move(str(src), str(target))
        else:
            shutil.copy2(str(src), str(target))
        landed.append(target)
    return landed


def build_consolidated(report, events=None, confirm_overwrite=None, force=False):
    """Build (or reuse) the report's consolidated/normalized Excel from its raw.
    Reuses the existing workbook when it is already current and not `force`.
    Resolves the report's builder lazily (so pdfplumber/openpyxl load only here)
    and delegates to it with the library's raw/out paths — the consolidator is
    untouched. Returns a ConsolidateResult."""
    spec = get(report)
    if not force and status(report)["current"]:
        return ConsolidateResult(
            status="ok",
            message="TSN consolidated workbook is already current; reused.",
            summary_lines=[f"{spec.label}: consolidated is up to date (reused)."],
            completion=outcome.COMPLETE,
            skipped_inputs=0,
            failed_inputs=0,
        )
    raws = _raw_files(report)
    if not raws:
        return ConsolidateResult(
            status="error",
            message=(f"No raw {spec.label} files in:\n{raw_dir(report)}\n\n"
                     "Import the raw TSN file(s) first (Settings ▸ TSN reports)."),
        )
    try:
        before_manifest = _raw_manifest(report, raws)
    except (OSError, ValueError) as e:
        return ConsolidateResult(
            status="error",
            message=(f"{spec.label}: the raw source could not be bound before "
                     f"normalization ({type(e).__name__}: {e})."),
        )
    mod_name, func_name = spec.builder.split(":")
    builder = getattr(importlib.import_module(mod_name), func_name)
    out = consolidated_path(report)
    out.parent.mkdir(parents=True, exist_ok=True)

    def uncertified(message):
        """Make any already-written unstable generation durably non-reusable."""
        marker = ConsolidateResult(
            status="ok", output_path=str(out), completion=outcome.PARTIAL,
            failed_inputs=1)
        consolidation_meta.write_outcome(
            out, marker,
            extra={"tsn_normalization_version": spec.normalization_version,
                   "tsn_raw_manifest": None,
                   "tsn_normalized_workbook_identity": None,
                   "tsn_artifact_identity_token": None})
        return ConsolidateResult(status="error", message=message)

    result = builder(raw_dir(report), out, events=events,
                     confirm_overwrite=confirm_overwrite)
    raw_manifest = None
    workbook_identity = None
    identity_token = None
    if result.status == "ok":
        try:
            after_manifest = _raw_manifest(report)
        except (OSError, ValueError) as e:
            return uncertified(
                f"{spec.label}: the raw source could not be rebound after "
                f"normalization ({type(e).__name__}: {e}); the workbook is not "
                "certified for reuse.")
        if after_manifest != before_manifest:
            return uncertified(
                f"{spec.label}: the raw member names or bytes changed during "
                "normalization; the workbook is not certified for reuse.")
        builder_manifest = getattr(result, "tsn_raw_manifest", None)
        try:
            builder_manifest = _raw_contract.validate_raw_manifest(builder_manifest)
        except ValueError as e:
            return uncertified(
                f"{spec.label}: the builder did not return a valid raw-source "
                f"certificate ({e}).")
        if builder_manifest != before_manifest:
            return uncertified(
                f"{spec.label}: the builder's source certificate does not match "
                "the canonical raw manifest.")
        raw_manifest = builder_manifest
        try:
            workbook_identity = normalized_workbook_identity(out)
            identity_token = canonical_normalized_identity_token(
                report, raw_manifest, workbook_identity)
        except ValueError as e:
            return uncertified(
                f"{spec.label}: the normalized workbook bytes could not be "
                f"bound after normalization ({e}); the workbook is not "
                "certified for reuse.")
    # P1-B05: persist the builder's producer completion beside the generated workbook
    # through the shared boundary, so a PARTIAL TSN normalization (categories / district
    # PDFs left out) stays flagged when resolve() reuses the consolidated workbook. A
    # False return = a non-complete normalization's flag could NOT be recorded
    # (publication failed): the build cannot claim a safely persisted artifact, so report
    # an error result rather than the success-shaped one (the worker/UI surfaces it).
    published = consolidation_meta.write_outcome(
        out, result,
        # The producer's own claims (e.g. the summaries' tsn_source_claims,
        # CMP-AUD-144/145/146) merge in additively; the library's binding keys
        # stay authoritative on collision.
        extra={**(getattr(result, "producer_extra", None) or {}),
               "tsn_normalization_version": spec.normalization_version,
               "tsn_raw_manifest": raw_manifest,
               "tsn_normalized_workbook_identity": workbook_identity,
               "tsn_artifact_identity_token": identity_token})
    if not published:
        return ConsolidateResult(
            status="error",
            message=(f"{spec.label}: normalization finished but its required outcome "
                     "and raw-source certificate could not be recorded; the workbook "
                     "is not certified for comparison or reuse — re-run."))
    # Generic consolidation metadata deliberately permits a COMPLETE workbook to
    # exist without a sidecar for legacy producers. TSN cannot: its raw manifest
    # and normalizer version are mandatory provenance. Re-read through the same
    # production status boundary consumers use; a helper-level True is not enough.
    certified = status(report)
    if not (certified.get("metadata_current")
            and certified.get("normalization_current")
            and certified.get("raw_manifest_current")
            and certified.get("normalized_workbook_current")
            and certified.get("identity_token_current")
            and certified.get("normalized_workbook_identity") == workbook_identity
            and certified.get("certified_identity_token") == identity_token):
        return ConsolidateResult(
            status="error",
            message=(f"{spec.label}: normalization finished but the durable TSN "
                     "source/workbook certificate could not be verified against the "
                     "exact normalized bytes. The produced workbook is stale/untrusted "
                     "and will not be used; re-run after the destination is writable."))
    result.tsn_normalized_workbook_identity = workbook_identity
    result.tsn_artifact_identity_token = identity_token
    return result


def ensure_current(report, events=None, source=None):
    """Auto-heal hook the compare paths call before READING the library's
    consolidated workbook (D2): a stale library — normalizer version mismatch or
    raw newer than the build — rebuilds itself from the retained raw, announced.
    Returns the rebuild's ConsolidateResult, or None only when nothing needs
    doing (current / unregistered / nothing built yet). If a consolidated
    workbook exists but its raw source is absent, unreadable, ambiguous, or
    otherwise uncertifiable, returns a typed error result so comparison callers
    stop instead of reading stale bytes. The no-consolidated/no-raw state remains
    None so the existing import-and-build UX is preserved.
    """
    if not is_registered(report):
        return None
    spec = get(report)
    if source is not None and source.get("kind") == "consolidated":
        resolved_path = source.get("path")
        try:
            is_canonical = bool(
                resolved_path
                and Path(resolved_path).absolute() == consolidated_path(report).absolute())
        except (OSError, RuntimeError, TypeError) as e:
            # Fail closed, but the refusal below only says "legacy or foreign"; the
            # actual path fault has to reach the log.
            log.warning("tsn: could not canonicalize the resolved consolidated path "
                        "%r for %s (%s: %s)", resolved_path, report,
                        type(e).__name__, e)
            is_canonical = False
        if source.get("legacy") or not is_canonical:
            return ConsolidateResult(
                status="error", completion=outcome.FAILED, failed_inputs=1,
                message=(f"{spec.label}: the resolved consolidated TSN workbook "
                         "is a legacy or foreign artifact with no canonical raw-source "
                         "certificate. Comparison was stopped. Import the authoritative "
                         "raw TSN source into Settings > TSN reports and rebuild it; "
                         "the uncertified workbook was not used."))
    st = status(report)
    if st["current"] or not st["consolidated_present"]:
        return None
    if not st.get("raw_admissible"):
        detail = (st.get("raw_admission_error")
                  or "the retained raw source cannot be certified")
        message = (
            f"{spec.label}: the existing consolidated TSN workbook is not "
            f"certifiably current, and comparison was stopped. It was not used "
            f"because {detail}. Restore the authoritative raw source in\n"
            f"{raw_dir(report)}\nthen rebuild it in Settings > TSN reports.")
        log.error("tsn library: refusing stale consolidated %s (%s)", report, detail)
        return ConsolidateResult(
            status="error", message=message, completion=outcome.FAILED,
            failed_inputs=1)

    ev = events or Events()
    why = ("normalization updated" if not st.get("normalization_current")
           else "raw files changed")
    ev.on_log(f"TSN library: rebuilding {spec.label} ({why})…")
    log.info("tsn library: auto-rebuild %s (%s)", report, why)
    result = build_consolidated(report, events=events)
    if result.status != "ok":
        return result
    certified = status(report)
    if not certified.get("current"):
        detail = (certified.get("raw_admission_error")
                  or "the rebuilt workbook has no complete durable source certificate")
        return ConsolidateResult(
            status="error", completion=outcome.FAILED, failed_inputs=1,
            message=(f"{spec.label}: the TSN rebuild returned success but the "
                     f"workbook is still not certifiably current ({detail}). "
                     "Comparison was stopped; rebuild it in Settings > TSN reports."))
    return result


# --------------------------------------------------------------------------- #
# Shared single-file normalizer (the tsn_load_* family substrate, S04/R1-N01)
#
# The four single-file TSN loaders (tsn_load_ramp_summary / _ramp_detail /
# _intersection_summary / _intersection_detail) all share ONE skeleton: find the
# sole ordinary raw export in raw/, parse it via the report's own projector (which lives in
# the matching compare_*_tsn module), then write a small normalized write-only
# workbook (one sheet, a styled header, the projected rows) atomically. Only the
# glob, projection, header, and result text differ per report — that per-report glue
# stays in the thin tsn_load_* shim; this factory owns the rest. compare_core is
# untouched and the builder strings (tsn_load_*:build_into) are preserved.
# --------------------------------------------------------------------------- #
def _statewide_raw_candidates(rdir, glob):
    """Sorted ordinary matching files for a single-statewide source.

    Excel owner-lock files are not source candidates.  No candidate wins by
    name or mtime: the caller must prove cardinality is exactly one.
    """
    return sorted((p for p in Path(rdir).glob(glob)
                   if p.is_file() and not p.name.startswith("~$")),
                  key=lambda p: p.name.casefold())


def _write_normalized_workbook(sheet, header, header_align, rows):
    """A write-only workbook: one `sheet`, the styled `header` row (the shared
    TSN-library blue header + the report's own Alignment kwargs), then `rows`."""
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet)
    head_fill = PatternFill("solid", start_color="305496")
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    align = Alignment(**header_align)
    cells = []
    for label in header:
        c = WriteOnlyCell(ws, value=label)
        c.fill, c.font, c.alignment = head_fill, head_font, align
        cells.append(c)
    ws.append(cells)
    for r in rows:
        ws.append(r)
    return wb


def build_normalized(raw_dir, out_path, *, glob, deps_ok, deps_msg, no_raw_what,
                     no_raw_hint, log_label, sheet, header, header_align, project,
                     events=None, confirm_overwrite=None, marker_version=None):
    """Shared driver for the single-file TSN normalizers (S04). Requires exactly
    one ordinary `glob` file in `raw_dir`, parses it via `project`, and writes the normalized
    workbook (one `sheet`, the styled `header`, then the projected rows) to
    `out_path` atomically (F9: temp + os.replace, never truncating a prior file).

    `project(raw_path)` returns `(rows, make_result)`:
      * `rows`                 -- the iterable of data rows appended under `header`;
      * `make_result(out_name)` -- builds the success ConsolidateResult, so each
        report keeps its own message / summary_lines / producer completion.
    The deps gate (`deps_ok`/`deps_msg`), missing-raw message (`No raw {no_raw_what}
    found ...` + `no_raw_hint`), overwrite-confirm, parse-error wrapping, and the
    PermissionError save guard are identical across the family and live here.
    Returns a ConsolidateResult; `compare_core` is untouched."""
    events = events or Events()
    if not deps_ok:
        return ConsolidateResult(status="error", message=deps_msg)
    candidates = _statewide_raw_candidates(raw_dir, glob)
    if not candidates:
        return ConsolidateResult(
            status="error",
            message=f"No raw {no_raw_what} found in:\n{raw_dir}\n\n{no_raw_hint}")
    if len(candidates) != 1:
        names = ", ".join(p.name for p in candidates)
        return ConsolidateResult(
            status="error",
            message=(f"Expected exactly one raw {no_raw_what} in:\n{raw_dir}\n\n"
                     f"Found {len(candidates)} ordinary matching files: {names}\n\n"
                     "Remove or move the extra source file(s), then rebuild."))
    raw = candidates[0]
    raw_root = Path(raw_dir)
    out_path = Path(out_path)
    confirm = confirm_overwrite or (lambda _p: True)
    if out_path.exists() and not confirm(out_path):
        return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    # Bind the exact bytes before any parser opens the source.  The parser consumes
    # a private snapshot with the same basename/extension, so even a transient
    # A->B->A rewrite wholly inside a long parse cannot create a mixed generation.
    try:
        source_manifest, captured = _raw_contract.capture_raw_manifest([raw], raw_root)
    except (OSError, ValueError) as e:
        return ConsolidateResult(
            status="error",
            message=(f"Could not bind {raw.name} before normalization: "
                     f"{type(e).__name__}: {e}"))
    relative = source_manifest["members"][0]["relative_path"]
    source_problem = [None]

    def source_current():
        try:
            current_candidates = _statewide_raw_candidates(raw_root, glob)
            if len(current_candidates) != 1:
                source_problem[0] = (
                    f"the raw source universe changed to {len(current_candidates)} "
                    "ordinary matching files")
                return False
            current = _raw_contract.canonical_raw_manifest(
                current_candidates, raw_root)
        except (OSError, ValueError) as e:
            source_problem[0] = f"{type(e).__name__}: {e}"
            return False
        if current != source_manifest:
            source_problem[0] = "the raw member name or bytes changed"
            return False
        return True

    def changed_result(previous_kept=True):
        detail = source_problem[0] or "the raw source changed"
        disposition = (
            "the previous normalized workbook was kept"
            if previous_kept else
            "the just-written snapshot was not certified as current and must not be reused"
        )
        return ConsolidateResult(
            status="error",
            message=(f"The {log_label} raw source changed while it was being "
                     f"normalized ({detail}); {disposition}."))

    events.on_log(f"Normalizing {log_label}: {raw.name}")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with tempfile.TemporaryDirectory(
                prefix=".tsn-statewide-source-", dir=out_path.parent) as snapshot_dir:
            snapshot = Path(snapshot_dir) / Path(relative).name
            snapshot.write_bytes(captured[relative])
            rows, make_result = project(str(snapshot))
    except Exception as e:
        return ConsolidateResult(
            status="error",
            message=f"Could not read {raw.name}: {type(e).__name__}: {e}")

    # Detect a persistent mutation before spending time serializing the workbook.
    # The same predicate runs again from atomic_save_if after serialization and
    # immediately before os.replace, preserving last-good on detected pre-commit drift.
    if not source_current():
        return changed_result()

    # A statewide export that parses but projects ZERO rows is a layout change
    # (a renamed sheet / moved header), not a legitimate empty dataset — writing
    # an "ok" header-only workbook here silently turned EVERY comparison row into
    # "Only in TSMIS". Error out; the previous normalized file stays in place.
    rows = list(rows)
    if not rows:
        return ConsolidateResult(
            status="error",
            message=(f"{raw.name} parsed but produced 0 rows — the {log_label} "
                     "layout may have changed. Nothing was written; the previous "
                     "normalized file (if any) was kept."))

    try:
        wb = _write_normalized_workbook(sheet, header, header_align, rows)
        if marker_version is not None:
            # CMP-AUD-037: the XLSX-sourced families (RD/ID/HD) stamp their
            # normalized workbook with the version so the DIRECT comparison path
            # can refuse a stale library (the matrix/library path already gates
            # via the certificate, D2). create_sheet + append works on the
            # write-only workbook.
            import compare_tsn_common as _ctc
            _ctc.write_normalization_marker(wb, marker_version, report_name=log_label)
    except ImportError:
        # The `deps_ok` probe is a single `from openpyxl import Workbook`; this is the
        # centralized backstop for a partial/frozen-pruned openpyxl whose WriteOnlyCell /
        # styles symbols are missing — return the shim's friendly deps message, never crash
        # (P5-A01). Kept here (not 4 broadened probes) so the writing skeleton stays in one place.
        return ConsolidateResult(status="error", message=deps_msg)
    import artifact_store
    try:
        committed = artifact_store.atomic_save_if(wb, out_path, source_current)
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not save {out_path.name}.\n\n"
                     "The file is probably open in Excel. Close it and try again."))
    if not committed:
        return changed_result()
    # There is an irreducible external-writer window between proceed() and the
    # filesystem replace. A post-replace recheck prevents direct builders from
    # returning false success if mutation lands in that window. The workbook is
    # still one truthful immutable snapshot (never a mixed read), but may have
    # replaced prior bytes; the library wrapper will keep it non-current.
    if not source_current():
        return changed_result(previous_kept=False)
    result = make_result(out_path.name)
    # build_consolidated requires this exact certificate for every registered
    # builder before it can persist a reusable normalization generation.
    result.tsn_raw_manifest = source_manifest
    return result


# --------------------------------------------------------------------------- #
# Resolve (the matrices' single TSN entry point)
# --------------------------------------------------------------------------- #
_SELECTION_VERSION = 1
_TSN_DATASET_ALIASES = {
    "highway_log_pdf": "highway_log",
    "intersection_detail_pdf": "intersection_detail",
    "highway_detail_pdf": "highway_detail",
    "highway_sequence_pdf": "highway_sequence",
    "ramp_detail_pdf": "ramp_detail",
}


def canonical_dataset_key(key):
    """The one TSN-library key for an export row or its PDF sibling."""
    key = str(key or "").strip()
    return _TSN_DATASET_ALIASES.get(key, key)


def selection_path(selection):
    """The display path from a versioned selection record or legacy string."""
    if isinstance(selection, str):
        return selection.strip() or None
    if isinstance(selection, dict):
        path = selection.get("path")
        return path.strip() if isinstance(path, str) and path.strip() else None
    return None


def _normalize_selection(selection):
    path = selection_path(selection)
    if not path:
        return None
    if isinstance(selection, str):
        return {"version": 0, "path": path}
    out = dict(selection)
    out["path"] = path
    if not isinstance(out.get("identity"), dict):
        out["identity"] = None
    return out


def canonicalize_selections(selections):
    """Return ``(canonical_map, changed)`` for persisted explicit selections.

    Historical PDF-export keys migrate to their shared base TSN dataset. Bare
    legacy paths become version-0 records and therefore require a deliberate
    re-pick. Conflicting aliases collapse to one fail-closed record instead of
    silently choosing whichever dict entry happened to be visited last.
    """
    source = selections if isinstance(selections, dict) else {}
    out = {}
    for raw_key, raw_selection in source.items():
        record = _normalize_selection(raw_selection)
        if record is None:
            continue
        key = canonical_dataset_key(raw_key)
        if key in out and out[key] != record:
            paths = sorted({p for p in (selection_path(out[key]), selection_path(record)) if p})
            out[key] = {"version": 0, "path": paths[0] if paths else "(conflict)",
                        "invalid_reason": "conflicting_aliases",
                        "legacy_paths": paths}
        else:
            out[key] = record
    return out, out != source


class _SelectionInspectionError(Exception):
    def __init__(self, reason, detail=""):
        super().__init__(detail)
        self.reason = reason
        self.detail = detail


def _stat_signature(st):
    return (int(st.st_size), int(getattr(st, "st_mtime_ns", st.st_mtime * 1e9)),
            int(getattr(st, "st_dev", 0)), int(getattr(st, "st_ino", 0)))


def _inspect_explicit_workbook(path):
    """Validate one XLSX and return a stable content + filesystem identity."""
    p = Path(path)
    try:
        before = p.stat()
        digest = hashlib.sha256()
        with p.open("rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(block)
        after_hash = p.stat()
    except FileNotFoundError as e:
        raise _SelectionInspectionError("missing", str(e)) from e
    except OSError as e:
        raise _SelectionInspectionError("unreadable", str(e)) from e
    if _stat_signature(before) != _stat_signature(after_hash):
        raise _SelectionInspectionError("changed_during_read")

    try:
        from openpyxl import load_workbook
        # Own the file handle so malformed workbook/XML failures cannot leave an
        # openpyxl ZIP handle locking the selected file on Windows.
        with p.open("rb") as workbook_stream:
            wb = load_workbook(workbook_stream, read_only=True, data_only=False,
                               keep_links=False)
            try:
                if not wb.sheetnames:
                    raise ValueError("workbook has no worksheets")
            finally:
                wb.close()
        after_open = p.stat()
    except OSError as e:
        raise _SelectionInspectionError("unreadable", str(e)) from e
    except Exception as e:  # noqa: BLE001 — malformed ZIP/XML is one not-workbook state
        raise _SelectionInspectionError("not_workbook", type(e).__name__) from e
    if _stat_signature(before) != _stat_signature(after_open):
        raise _SelectionInspectionError("changed_during_read")

    file_id = (f"{int(getattr(before, 'st_dev', 0))}:"
               f"{int(getattr(before, 'st_ino', 0))}")
    return {"sha256": digest.hexdigest(), "size": int(before.st_size),
            "mtime_ns": int(getattr(before, "st_mtime_ns", before.st_mtime * 1e9)),
            "file_id": file_id}


def create_explicit_selection(path):
    """Create the only trusted persisted explicit-selection record."""
    try:
        p = Path(path).expanduser().resolve()
    except (TypeError, OSError, RuntimeError) as e:
        raise ValueError("Pick a readable Excel workbook (.xlsx) for the TSN source.") from e
    if p.suffix.lower() != ".xlsx":
        raise ValueError("Pick an Excel workbook (.xlsx) for the TSN source.")
    try:
        identity = _inspect_explicit_workbook(p)
    except _SelectionInspectionError as e:
        if e.reason == "missing":
            msg = f"The selected TSN workbook no longer exists: {p}."
        elif e.reason == "not_workbook":
            msg = f"The selected TSN file is not a readable Excel workbook: {p}."
        elif e.reason == "changed_during_read":
            msg = f"The selected TSN workbook changed while it was being checked: {p}."
        else:
            msg = f"The selected TSN workbook could not be read: {p}."
        raise ValueError(msg) from e
    return {"version": _SELECTION_VERSION, "path": str(p), "identity": identity}


def _missing_explicit(record, reason):
    return {"kind": "missing_explicit", "selected_path": selection_path(record),
            "selection_reason": reason}


def resolve(report, selected_file=None, legacy_dest=None):
    """Resolve the TSN dataset for `report` (the matrices' single entry point) and
    attach the persisted producer `completion` (P1-B05) to any path-bearing source —
    so a PARTIAL generated TSN workbook (categories / district PDFs left out) flags the
    comparison even when reused. `completion` is None for a user-picked file or a
    workbook with no sidecar (deliberate: a user pick / legacy workbook reads complete)."""
    r = _resolve_source(report, selected_file, legacy_dest)
    if r.get("path"):
        r = {**r, "completion": consolidation_meta.read_completion(r["path"])}
    return r


def _resolve_source(report, selected_file=None, legacy_dest=None):
    """Resolve the TSN dataset for `report`, returning the same contract as
    matrix.tsn_source: {kind: file|consolidated|pdfs|raw|none, path?, mtime?,
    pdf_count?, raw_count?}. Resolution order:
      1. an explicit user-picked `selected_file` — a real .xlsx wins; a missing or
         invalid pick returns ``missing_explicit`` and FAILS CLOSED (it never falls
         through to a different canonical/legacy dataset);
      2. the library's generated consolidated workbook;
      3. the library's raw file(s) -> 'pdfs' (PDF raw) / 'raw' (xlsx raw): the
         'import is present but build the consolidated first' state;
      4. the dest-scoped legacy drop <dest>/_tsn_input/<report>/ (back-compat);
      5. the global legacy locations (Highway Log only);
      6. none.
    """
    if selected_file:
        record = _normalize_selection(selected_file)
        if record is None:
            return _missing_explicit({"path": "(unknown path)"}, "legacy_identity")
        if record.get("invalid_reason") == "conflicting_aliases":
            return _missing_explicit(record, "conflicting_aliases")
        p = Path(record["path"])
        if p.suffix.lower() != ".xlsx":
            return _missing_explicit(record, "not_xlsx")
        if record.get("version") != _SELECTION_VERSION or not isinstance(
                record.get("identity"), dict):
            return _missing_explicit(record, "legacy_identity")
        try:
            current_identity = _inspect_explicit_workbook(p)
        except _SelectionInspectionError as e:
            return _missing_explicit(record, e.reason)
        if current_identity != record["identity"]:
            return _missing_explicit(record, "changed")
        return {"kind": "file", "path": str(p), "mtime": _safe_mtime(p),
                "selected_path": str(p), "explicit": True,
                "selection": record, "identity_token": current_identity}

    if is_registered(report):
        cons = consolidated_path(report)
        if cons.is_file():
            certified = status(report)
            return {"kind": "consolidated", "path": str(cons),
                    "mtime": _safe_mtime(cons),
                    # Existing matrix cache plumbing already persists an optional
                    # identity token. Canonical library artifacts now expose the
                    # same content-bound surface as explicit picks, without making
                    # consumers reconstruct TSN provenance from loose fields.
                    # CMP-AUD-081: this is None whenever the library is not current
                    # (status: `identity_token if current else None`), so a
                    # would-rebuild library reads dependent Matrix cells stale.
                    "identity_token": certified.get("identity_token")}
        raws = _raw_files(report)
        if raws:
            spec = get(report)
            if spec.raw_kind in ("district_pdfs", "statewide_pdf"):
                return {"kind": "pdfs", "pdf_count": len(raws)}
            return {"kind": "raw", "raw_count": len(raws), "raw_kind": spec.raw_kind}

    if legacy_dest is not None:
        r = _resolve_dest_drop(legacy_dest, report)
        if r["kind"] != "none":
            return r

    r = _resolve_legacy_global(report)
    if r["kind"] != "none":
        return r

    return {"kind": "none"}


def explicit_selection_problem(source):
    """Actionable message when a persisted explicit TSN selection is unusable."""
    if not isinstance(source, dict) or source.get("kind") != "missing_explicit":
        return None
    picked = source.get("selected_path") or "(unknown path)"
    reason = source.get("selection_reason")
    why = {
        "not_xlsx": "is not an .xlsx workbook",
        "missing": "no longer exists",
        "unreadable": "cannot be read",
        "not_workbook": "is not a readable Excel workbook",
        "changed": "changed since it was selected",
        "changed_during_read": "changed while it was being checked",
        "legacy_identity": "was saved by an older version without a verifiable identity",
        "conflicting_aliases": "conflicts with another saved selection for the same TSN dataset",
    }.get(reason, "is unavailable")
    return (f"The selected TSN workbook {why}: {picked}. "
            "Re-pick that TSN file, or clear the selection to use the canonical library.")


def require_explicit_selection(selection):
    """Revalidate a queued explicit choice immediately around comparison work."""
    src = _resolve_source("", selection)
    if src.get("kind") != "file":
        raise ValueError(explicit_selection_problem(src))
    return src
