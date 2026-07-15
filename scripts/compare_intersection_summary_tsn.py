"""Build the TSMIS-vs-TSN Intersection Summary discrepancy workbook (AGGREGATE).

The Ramp Summary recipe applied to the intersection category taxonomy: each side
reduces to ONE statewide {category: count} table, compared with has_route=False
(key = category, field = count). The category schema (the UNION of both systems'
11 blocks — incl. the diverged CONTROL TYPES / INTERSECTION TYPE codes that show
one-sided) + the block-walk mapper + the familiar-layout sheet all live in
summary_layout (INTERSECTION_SUMMARY_SPEC), shared with the consolidator.

  * TSMIS side — the CONSOLIDATED Intersection Summary workbook (one row per route,
    one column per category key); summed column-by-column.
  * TSN side — the statewide PDF: a 3-COLUMN page (left/middle/right bands), each
    band block-walked with the SAME summary_layout.counts_from_rows the TSMIS
    consolidator uses, so the two sides can't drift.

Reconciled on the 6.19 ground truth: TSMIS 16473 vs TSN 16626. CONTROL TYPES: the
TSN signal sub-types J–P fold into the shared "Signalized" (S) category — the same
crosswalk the Detail uses (per the TSNR/MIRE reference), applied in
summary_layout.counts_from_rows for BOTH sides — so Signalized compares directly
(TSMIS S vs TSN J–P summed). The remaining codes the TSN summary genuinely doesn't
tabulate (CONTROL Roundabout R / PHB O / Flash Q; INTERSECTION TYPE R/C/P) stay
one-sided; the "+ no data" buckets the TSN PDF reports as 0 ARE compared. Console-free.
"""
import re
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from collections import Counter
from dataclasses import replace

import compare_tsn_common as ctc
import consolidation_meta
import summary_layout
from compare_core import CompareSchema
from paths import today_str

REPORT_NAME = "Intersection Summary"
TSMIS_SHEET = "Intersection Summary"            # consolidated per-route sheet
NORMALIZED_SHEET = "Intersection Summary (TSN)"  # library normalized 2-col workbook

_SPEC = summary_layout.INTERSECTION_SUMMARY_SPEC
_CATEGORIES = _SPEC.categories()                 # [(key, slug), ...] incl. Total
# Column-header (compare key) -> slug for reading the consolidated workbook.
_KEY_TO_SLUG = {c.key: c.slug for sec in _SPEC.sections for c in sec.cats}
_KEY_TO_SLUG[_SPEC.total.key] = _SPEC.total.slug

# 3-column band boundaries on the statewide page (calibrated to the raw geometry).
_BAND_LEFT, _BAND_RIGHT = 190, 495
_TOTAL_RE = re.compile(r"Total Intersections\s*=\s*([\d,]+)", re.IGNORECASE)

# Slug the folded "Signalized" category lands in (resolved from the spec, not hardcoded).
_SIGNALIZED_SLUG = next(
    (c.slug for sec in _SPEC.sections if sec.name == summary_layout._IS_CONTROL_TYPES
     for c in sec.cats if c.code == "S"), "is_control_types_s")
_STALE_SIGNAL_KEY = re.compile(r"CONTROL TYPES:\s*([A-Za-z])\b")

# The censused partition contract (CMP-AUD-020). TSMIS side: verified statewide
# on the 6.19 and 7.9 exports (mirrors the consolidator's per-route tripwire) —
# every block partitions the total exactly EXCEPT Highway Group, which the site
# itself under-counts. TSN side: censused on the 2025-09 statewide print — five
# blocks partition exactly; the blocks holding TSMIS-only codes (intersection
# R/C/P/+, control R/O/Q, left-chan Y) sum short by exactly the untabulated
# classes, the lane rows don't cover >8 lanes, and right-channelization leaves a
# small censused remainder (3 on that print). Bounded residuals are EXPOSED as
# notes, never fabricated into a category.
_RULE = summary_layout.SectionRule
_TSMIS_RULES = (
    _RULE("HIGHWAY GROUP", "bounded",
          reason="the TSMIS site under-counts the Highway Group block (a known "
                 "site-side tabulation gap; every other block partitions exactly)"),
    _RULE(summary_layout._IS_RURAL_URBAN, "exact"),
    _RULE("INTERSECTION TYPE", "exact"),
    _RULE("LIGHTING TYPE", "exact"),
    _RULE(summary_layout._IS_CONTROL_TYPES, "exact"),
    _RULE("MAINLINE NUM OF LANES", "exact"),
    _RULE("MAINLINE MASTARM", "exact"),
    _RULE("MAINLINE LEFT CHANNELIZATION", "exact"),
    _RULE("MAINLINE RIGHT CHANNELIZATION", "exact"),
    _RULE("MAINLINE TRAFFIC FLOW", "exact"),
)
_TSN_RULES = (
    _RULE("HIGHWAY GROUP", "exact"),
    _RULE(summary_layout._IS_RURAL_URBAN, "exact"),
    _RULE("INTERSECTION TYPE", "bounded",
          reason="the TSN summary doesn't tabulate the TSMIS-only intersection "
                 "types (Roundabout/Circular/Midblock/no-data); see 'Only in "
                 "TSMIS'"),
    _RULE("LIGHTING TYPE", "exact"),
    _RULE(summary_layout._IS_CONTROL_TYPES, "bounded",
          reason="the TSN summary doesn't tabulate the TSMIS-only control types "
                 "(Roundabout R / PHB O / Flash Beacon Q); see 'Only in TSMIS'"),
    _RULE("MAINLINE NUM OF LANES", "bounded",
          reason="the TSN print's lane rows don't cover every intersection "
                 "(lanes above 8 are not tabulated)"),
    _RULE("MAINLINE MASTARM", "exact"),
    _RULE("MAINLINE LEFT CHANNELIZATION", "bounded",
          reason="the TSN summary has no 'channelization not specified' (Y) row"),
    _RULE("MAINLINE RIGHT CHANNELIZATION", "bounded",
          reason="the TSN statewide print leaves a small censused remainder in "
                 "this block (3 intersections on the 2025-09 print)"),
    _RULE("MAINLINE TRAFFIC FLOW", "exact"),
)


def _slug_for_key(key_to_slug, key):
    """Resolve a normalized-workbook category key to its slug, tolerating a STALE
    library built before the J–P→Signalized fold (the library is REUSED, not rebuilt,
    after a code change — see tsn_library.build_consolidated). The legacy signal
    sub-type keys (CONTROL TYPES J–P) and the old 'S - SIGNALIZED' key all route to
    the folded Signalized slug, so a reused pre-fold library still compares correctly
    (their counts are summed)."""
    direct = key_to_slug.get(key)
    if direct is not None:
        return direct
    m = _STALE_SIGNAL_KEY.match(key)
    if m and m.group(1).upper() in summary_layout._CONTROL_SIGNAL_FOLD | {"S"}:
        return _SIGNALIZED_SLUG
    return None


_SCHEMA = CompareSchema(
    report_name=REPORT_NAME,
    header=["Category", "Count"],
    side_a="TSMIS",
    side_b="TSN",
    id_noun="category",
    id_noun_plural="categories",
    sides_noun="systems",
    cmp_widths={"Count": 12},
    data_widths={"Count": 12},
    scope_flat="Statewide category totals",
    one_sided_note_extra=" (a category one system classifies and the other doesn't)",
    extra_sheet_writer=summary_layout.make_extra_sheet_writer(_SPEC),
)


# --------------------------------------------------------------------------- #
# TSN side: 3-column statewide PDF -> {slug: count}
# --------------------------------------------------------------------------- #
def _cluster(words):
    """A band's words -> [(count_or_None, code-text)] rows (top-clustered, x-sorted)."""
    if not words:
        return []
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    rows, cur, ct = [], [], None
    for w in words:
        if ct is None or abs(w["top"] - ct) <= 3:
            cur.append(w)
            ct = ct if ct is not None else w["top"]
        else:
            rows.append(cur)
            cur, ct = [w], w["top"]
    if cur:
        rows.append(cur)
    out = []
    for row in rows:
        row.sort(key=lambda w: w["x0"])
        ts = [w["text"] for w in row]
        if ts and re.fullmatch(r"-?[\d,]+", ts[0]) and any(c.isdigit() for c in ts[0]):
            out.append((int(ts[0].replace(",", "")), " ".join(ts[1:])))
        else:
            out.append((None, " ".join(ts)))
    return out


def _data_page(pdf):
    for pg in pdf.pages:
        if _TOTAL_RE.search(pg.extract_text() or ""):
            return pg
    return pdf.pages[-1]


def parse_tsn_pdf(path):
    """The statewide TSN Intersection Summary PDF -> {slug: count}. Splits the page
    into 3 column bands and block-walks each independently (so each column's
    count+label rows pair correctly), merging the per-band counts."""
    name = Path(path).name
    with pdfplumber.open(path) as pdf:
        page = _data_page(pdf)
        bands = [[], [], []]
        for w in page.extract_words():
            x = w["x0"]
            bands[0 if x < _BAND_LEFT else (1 if x < _BAND_RIGHT else 2)].append(w)
        counts = {}
        for band in bands:
            for slug, v in summary_layout.counts_from_rows(
                    _SPEC, _cluster(band), source=name).items():
                counts[slug] = counts.get(slug, 0) + v
        m = _TOTAL_RE.search(page.extract_text() or "")
        counts["total_intersections"] = int(m.group(1).replace(",", "")) if m else None
    return counts


# --------------------------------------------------------------------------- #
# TSN source claims (CMP-AUD-144/145/146): the print's own words, preserved.
# --------------------------------------------------------------------------- #
# The censused raw CONTROL F/G descriptors (the 2025-09 statewide print): the
# print's F text erroneously repeats G's "(RED ON ALL)" meaning; the canonical
# F mapping (Red on Mainline) comes from the authoritative TSNR crosswalk. The
# mapping is a DECLARED CORRECTION, not an equality — if either printed label
# drifts from this census, refuse and re-census rather than silently remapping.
_RAW_CONTROL_F = "F-FOUR WAY FLASHER (RED ON ALL)"
_RAW_CONTROL_G = "G-FOUR WAY FLASHER (RED ON ALL)"
_TSNR_DECISION_SOURCE = "TSNR - Intersection Control and Geometry Type_4.25.24_AT 1.xlsx"


def _claims_from_rows(band_rows, full_text, source):
    """The print's source claims (CMP-AUD-144/145/146) from the clustered data
    rows + full document text: report identity, EVERY printed (block, count,
    raw-label) row pre-fold, the J–P signal components behind the derived
    Signalized count, and the declared CONTROL F correction. Refuses if the
    censused F/G descriptors drifted."""
    headers = {summary_layout._norm_header(s.name): s.name for s in _SPEC.sections}
    for s in _SPEC.sections:
        for alias in s.aliases:
            headers[summary_layout._norm_header(alias)] = s.name
    printed, cur = [], None
    for count, text in band_rows:
        t = str(text or "").strip()
        h = summary_layout._norm_header(t)
        if h in headers:
            cur = headers[h]
            continue
        if cur is not None and count is not None:
            printed.append([cur, int(count), t])
    control = {t.split("-", 1)[0].strip().upper(): (n, t)
               for blk, n, t in printed
               if blk == summary_layout._IS_CONTROL_TYPES and "-" in t}
    components = [[c, control[c][0]] for c in "JKLMNPS" if c in control
                  and c in summary_layout._CONTROL_SIGNAL_FOLD | {"S"}]
    for code, censused in (("F", _RAW_CONTROL_F), ("G", _RAW_CONTROL_G)):
        got = control.get(code, (None, None))[1]
        if got != censused:
            raise ValueError(
                f"{source}: the print's CONTROL {code} descriptor ({got!r}) no "
                f"longer matches the censused text ({censused!r}) — the source "
                "label changed; re-census the declared F/G correction before "
                "normalizing")
    return {
        "schema_version": 1,
        "identity": ctc.tsn_print_identity(full_text, source),
        "printed_rows": printed,
        "signal_components": components,
        "declared_corrections": [{
            "block": summary_layout._IS_CONTROL_TYPES, "code": "F",
            "printed": _RAW_CONTROL_F,
            "canonical": "4-WAY FLASHER (RED/MAINLINE)",
            "decision_source": _TSNR_DECISION_SOURCE,
        }],
    }


def parse_tsn_source_claims(path):
    """The statewide TSN PDF's source claims (see _claims_from_rows)."""
    path = Path(path)
    try:
        with pdfplumber.open(path) as pdf:
            full_text = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
            page = _data_page(pdf)
            bands = [[], [], []]
            for w in page.extract_words():
                x = w["x0"]
                bands[0 if x < _BAND_LEFT else (1 if x < _BAND_RIGHT else 2)].append(w)
            rows = [r for band in bands for r in _cluster(band)]
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Could not read {path.name}: {type(e).__name__}: {e}")
    return _claims_from_rows(rows, full_text, path.name)


def validate_claims_against_counts(claims, counts):
    """CMP-AUD-144: the derived Signalized count must equal the sum of the
    print's own J–P (+ any printed S) component claims, exactly."""
    components = claims.get("signal_components") or []
    total = sum(n for _c, n in components)
    derived = counts.get(_SIGNALIZED_SLUG)
    if components and derived != total:
        raise ValueError(
            f"the derived Signalized count ({derived}) does not equal the sum "
            f"of the print's J–P/S component claims ({total}) — the fold and "
            "the source claims disagree")


def claims_notes(claims, side_label="TSN"):
    """Human-readable exposure lines for the familiar sheet + log."""
    if not claims:
        return [f"{side_label} print: no source-claims record beside this "
                "normalized workbook (older normalization) — rebuild the TSN "
                "library to capture the print identity."]
    ident = claims.get("identity") or {}
    notes = []
    if ident:
        notes.append(
            f"{side_label} print identity: {ident.get('report_id')} · Event "
            f"{ident.get('event_id')} · reference {ident.get('reference_date')} "
            f"· submitted by {ident.get('submitter')} · generated "
            f"{ident.get('generated_time')} ({ident.get('location_criteria')}).")
    comps = claims.get("signal_components") or []
    if comps:
        notes.append(
            f"{side_label} Signalized (S) is derived: the print's "
            + " + ".join(f"{c} {n:,}" for c, n in comps)
            + f" = {sum(n for _c, n in comps):,} signal sub-type records.")
    for corr in claims.get("declared_corrections") or ():
        notes.append(
            f"Declared correction: the {side_label} print labels {corr['block']} "
            f"{corr['code']} {corr['printed']!r} (a known print defect); the "
            f"canonical meaning is {corr['canonical']!r} per {corr['decision_source']}.")
    return notes


def _load_tsn(path):
    """TSN side -> {slug: count}. Raw statewide PDF, or the library's normalized
    Category|Count workbook."""
    path = Path(path)
    if path.suffix.lower() == ".pdf":
        # ValueError-on-unreadable, like the XLSX branch below (the shared
        # comparator catches ValueError and reports it cleanly).
        try:
            return parse_tsn_pdf(path)
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"Could not read {path.name}: {type(e).__name__}: {e}")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {path.name}: {type(e).__name__}: {e}")
    try:
        sn = NORMALIZED_SHEET if NORMALIZED_SHEET in wb.sheetnames else wb.sheetnames[0]
        key_to_slug = {k: s for k, s in _CATEGORIES}
        rec, seen = {}, set()
        it = wb[sn].iter_rows(values_only=True)
        next(it, None)
        for r in it:
            if not r or r[0] is None:
                continue
            key = str(r[0]).strip()
            slug = _slug_for_key(key_to_slug, key)
            if slug is None:
                continue
            if key in seen:                              # CMP-AUD-022
                raise ValueError(f"{path.name} lists the category {key!r} twice "
                                 "— refusing an ambiguous normalized table")
            seen.add(key)
            v = r[1] if len(r) > 1 else None
            if v is None:
                raise ValueError(f"{path.name}: the category {key!r} has no count")
            # Sum, not overwrite: DISTINCT stale J–P/S keys legitimately fold
            # into the one Signalized slug; a repeated EXACT key refused above.
            rec[slug] = rec.get(slug, 0) + summary_layout.parse_count(
                v, source=path.name, category=key)
        return rec
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# TSMIS side: SUM the consolidated workbook's per-route columns -> {slug: count}
# --------------------------------------------------------------------------- #
def _validate_route_universe(routes, name, workbook_path, note_sink):
    """CMP-AUD-183: the aggregated route universe must be sound — every data row
    carries a usable route identity and no route appears twice — and, when the
    producer recorded its ordered `route_census` in the outcome sidecar, must
    match that census EXACTLY (a dropped, added, renamed, reordered, or
    suffix-collapsed row refuses). A census-less workbook (older consolidation,
    or a copy without its sidecar) keeps the internal checks and gets an
    explicit no-census diagnostic instead of a silent pass."""
    cleaned = []
    for row_no, rv in routes:
        s = ("" if rv is None else str(rv)).strip()
        if isinstance(rv, bool) or not s or not s.isalnum() or len(s) > 8:
            raise ValueError(f"{name}: sheet row {row_no} has no usable route "
                             f"identity ({rv!r}) — its counts cannot be attributed")
        cleaned.append(s)
    if not cleaned:
        raise ValueError(f"{name} has no route rows — nothing to aggregate")
    dupes = sorted(r for r, n in Counter(cleaned).items() if n > 1)
    if dupes:
        raise ValueError(f"{name}: route(s) {', '.join(dupes[:6])} appear on more "
                         "than one row — refusing to aggregate an ambiguous "
                         "route universe")
    census = consolidation_meta.read_extra(workbook_path, "route_census")
    if census is None:
        note = (f"TSMIS route universe: {len(cleaned)} routes; no producer route "
                "census recorded (older consolidation) — internal checks only.")
    elif (not isinstance(census, list)
          or not all(isinstance(r, str) and r for r in census)):
        raise ValueError(f"{name}: the producer route census beside the workbook "
                         "is malformed — re-consolidate before comparing")
    elif census != cleaned:
        detail = (f"workbook has {len(cleaned)} route rows, the census records "
                  f"{len(census)}")
        for i, (want, got) in enumerate(zip(census, cleaned)):
            if want != got:
                detail = (f"first divergence at row {i + 2}: census {want!r} vs "
                          f"workbook {got!r}")
                break
        raise ValueError(f"{name}: the aggregated routes do not match the "
                         f"producer's route census ({detail}) — a route was "
                         "dropped, added, renamed, or reordered after "
                         "consolidation; re-consolidate before comparing")
    else:
        note = (f"TSMIS route universe verified against the producer census: "
                f"{len(cleaned)} routes ({cleaned[0]}–{cleaned[-1]}).")
    if note_sink is not None:
        note_sink.append(note)


def _load_tsmis(path, note_sink=None):
    """TSMIS side -> {slug: count}. Sums each category column of the consolidated
    Intersection Summary workbook's per-route sheet. Strict (CMP-AUD-021/022):
    count cells must be whole numbers (numeric text is parsed, never dropped;
    fractions and booleans refuse), a duplicated category column refuses, and a
    column the workbook lacks stays ABSENT — never a fabricated zero
    (reconcile_counts decides whether it was required). The route universe is
    validated per CMP-AUD-183 (see _validate_route_universe); the census status
    line lands in `note_sink`."""
    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        if TSMIS_SHEET not in wb.sheetnames:
            raise ValueError(f"{name} has no '{TSMIS_SHEET}' sheet — pick the "
                             "consolidated TSMIS Intersection Summary workbook.")
        it = wb[TSMIS_SHEET].iter_rows(values_only=True)
        header = next(it, None) or ()
        names = [str(h).strip() if h is not None else None for h in header]
        if "Route" not in names:
            raise ValueError(f"{name} isn't a consolidated Intersection Summary "
                             "workbook (no 'Route' column) — consolidate first.")
        route_col = names.index("Route")
        col_slug, seen = {}, set()
        for i, h in enumerate(names):
            if not h or h not in _KEY_TO_SLUG:
                continue
            if h in seen:                                # CMP-AUD-022
                raise ValueError(f"{name} has a duplicated category column "
                                 f"({h!r}) — refusing to sum an ambiguous table")
            seen.add(h)
            col_slug[i] = (_KEY_TO_SLUG[h], h)
        sums = {slug: 0 for slug, _h in col_slug.values()}
        routes = []
        for row_no, row in enumerate(it, start=2):
            if not row or all(c is None for c in row):
                continue
            routes.append((row_no, row[route_col] if route_col < len(row) else None))
            for i, (slug, disp) in col_slug.items():
                v = row[i] if i < len(row) else None
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue                             # blank cell
                sums[slug] += summary_layout.parse_count(v, source=name,
                                                         category=disp)
        _validate_route_universe(routes, name, path, note_sink)
        return sums
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# rows + adapter surface
# --------------------------------------------------------------------------- #
def _rows(counts, side):
    """Rows for `side` ('tsmis'|'tsn'): the side's applicable categories only, so a
    category the other system doesn't classify stays ONE-SIDED (Only in …).
    Every emitted category must be PRESENT in `counts` (reconcile_counts
    guarantees it) — an absent category is a hard error, never a fabricated
    zero (CMP-AUD-020/021)."""
    out = []
    for key, slug in _SPEC.categories_for(side):
        if slug not in counts:
            raise ValueError(f"the {side.upper()} table is missing the "
                             f"category {key!r} — cannot build comparison rows")
        out.append([key, counts[slug]])
    return out


def suggest_name(tsmis_path):
    return f"TSMIS_vs_TSN_IntersectionSummary_Comparison {today_str()}.xlsx"


def _load_pair(tsmis_path, tsn_path, note_sink=None, events=None):
    """(rows_t, rows_n, warnings) for the shared driver. Both inputs are
    INDEPENDENTLY validated before any row is built (CMP-AUD-020): every
    side-applicable category and the grand total must be present (the
    TSMIS-only codes are legitimately absent from the TSN PDF and are NOT
    required there), and every block must partition the total per the censused
    SectionRule contract — a table that doesn't reconcile refuses with a named
    block instead of comparing garbage. The TSMIS route universe is validated
    and reconciled against the producer's route census (CMP-AUD-183). Bounded
    residuals (the TSN-untabulated classes; the site's Highway Group
    under-count) and the census status are EXPOSED via `note_sink` onto the
    familiar sheet + the log — never fabricated into a category, and never a
    warning (warnings mean unreadable inputs)."""
    notes = []
    tsmis_counts = _load_tsmis(tsmis_path, note_sink=notes)   # + route census status
    tsn_counts = _load_tsn(tsn_path)
    # CMP-AUD-144/145/146: the print's own claims — fresh from a raw PDF (and
    # cross-checked against the folded counts), or the normalization sidecar's
    # record for a library workbook (absent -> explicit diagnostic note).
    if Path(tsn_path).suffix.lower() == ".pdf":
        claims = parse_tsn_source_claims(tsn_path)
        validate_claims_against_counts(claims, tsn_counts)
    else:
        claims = consolidation_meta.read_extra(tsn_path, "tsn_source_claims")
    notes += claims_notes(claims)
    notes += summary_layout.reconcile_counts(
        _SPEC, tsmis_counts, "tsmis", _TSMIS_RULES,
        source=Path(tsmis_path).name, side_label="TSMIS")
    notes += summary_layout.reconcile_counts(
        _SPEC, tsn_counts, "tsn", _TSN_RULES,
        source=Path(tsn_path).name, side_label="TSN")
    if note_sink is not None:
        note_sink.clear()
        note_sink.extend(notes)
    if events is not None:
        for n in notes:
            events.on_log(f"note: {n}")
    return _rows(tsmis_counts, "tsmis"), _rows(tsn_counts, "tsn"), []


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Intersection Summary TSMIS-vs-TSN AGGREGATE comparison workbook(s)."""
    # CMP-AUD-020: a per-run holder the loader fills and the familiar-sheet
    # writer reads, so censused-residual notes reach the display sheet without
    # ever entering the compared universe.
    notes = []
    schema = replace(_SCHEMA, extra_sheet_writer=summary_layout.make_extra_sheet_writer(
        _SPEC, extra_notes=notes))
    return ctc.run_files_compare(
        schema, tsmis_path, tsn_path, out_path,
        banner="Intersection Summary Comparison — TSMIS vs TSN (statewide category counts)",
        has_route=False,
        loader=lambda a, b: _load_pair(a, b, note_sink=notes, events=events),
        deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (pdfplumber, openpyxl).",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
