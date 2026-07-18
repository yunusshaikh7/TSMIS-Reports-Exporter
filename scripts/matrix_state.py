"""The matrix STATE side (S4 / ARC-03, split from matrix.py).

Pure filesystem READS: cell paths, the results sidecars, read_counts, the
shared `_staleness` reader behind comparison_state/_cmp_state, the TSN source
plumbing, and `matrix_snapshot` (what the GUI renders). No builds here — the
build side lives in matrix_build; `matrix` is the facade both re-export
through (patch `matrix.<name>` and every caller sees it).
"""
import json
import logging
import os
import time
from pathlib import Path

import artifact_store
import cache_envelope
import consolidation_meta
import outcome
import report_library
import reports
from common import DATA_SOURCES, ENVIRONMENTS

log = logging.getLogger("tsmis.matrix")

BASELINE_DEFAULT = "ssor-prod"
COMPARISONS_DIRNAME = "comparisons"
_RESULTS_FILE = "_results.json"
_MTIME_TOL_S = consolidation_meta._MTIME_TOL_S   # single home: the sidecar layer owns the tolerance


def producer_identity():
    """The semantic PRODUCER version a fresh comparison cache record is stamped
    with — the app's released ``MAJOR.MINOR.PATCH`` (CMP-AUD-084).

    A shipped comparator / parser / normalizer / consolidator fix always rides a
    new release, so binding every cache record to the app version invalidates any
    cell built by an OLDER pipeline exactly once on upgrade — a rebuild that
    recomputes against the current code, never a silent stale verdict — while an
    unchanged version reads fresh (a no-op within a release). This is the ONGOING
    semantic gate; the ``cache_envelope`` schema version stays the SEPARATE
    record-SHAPE migration (the finding's "separate record-shape migration from
    semantic invalidation"). A dict so a future finer per-producer key is an
    additive mismatch (stale-once), not a shape migration. Kept in ONE place so the
    Everything / by-day / baseline caches all agree on what "current" means, and
    routed through ``consolidation_meta.producer_app_version`` so the comparison-cache
    and consolidation freshness gates read the SAME value (robust to a scripts-only
    ``sys.path`` in isolated checks)."""
    return {"app": consolidation_meta.producer_app_version()}


# --------------------------------------------------------------------------- #
# cell identity
# --------------------------------------------------------------------------- #
def env_keys():
    """The matrix columns: every data-source/environment combo, in display order
    (ssor before ars; prod/test/dev)."""
    return [f"{s}-{e}" for s in DATA_SOURCES for e in ENVIRONMENTS]


def default_env_label(key):
    """A readable column label from an env key, with no dependency on the report
    registry: 'ssor-prod' -> 'SSOR / Prod'."""
    src, _, env = str(key).partition("-")
    return f"{src.upper()} / {env.title()}" if env else str(key).upper()


def _row_defs():
    """{row_key: (label, subdir, export_idx, adapter, has_route)} from the
    registry. has_route: the XLSX reports compare in the consolidated (Route +
    columns) shape; Ramp Summary (PDF, no sheet_name) is per-route."""
    out = {}
    for row_key, label, subdir, export_idx, adapter in reports.matrix_rows():
        has_route = getattr(adapter, "sheet_name", None) is not None
        out[row_key] = (label, subdir, export_idx, adapter, has_route)
    return out


# --------------------------------------------------------------------------- #
# paths (decision b: <dest>/comparisons/<baseline>/...) — STABLE, dateless names
# --------------------------------------------------------------------------- #
def comparisons_root(dest, baseline_key):
    return Path(dest) / COMPARISONS_DIRNAME / baseline_key


def comparison_path(dest, baseline_key, row_key, cell_key):
    """The comparison workbook for (row, cell-env) under one baseline.

    DELIBERATELY a deterministic, dateless name (NOT adapter.suggest_name, which
    embeds today's date): the matrix overwrites in place and uses the file's
    MTIME as the freshness signal, so the name must be stable across days. Do not
    'fix' this to a dated name."""
    return comparisons_root(dest, baseline_key) / f"{cell_key}_{row_key}.xlsx"


def _results_path(dest, baseline_key):
    return comparisons_root(dest, baseline_key) / _RESULTS_FILE


# --------------------------------------------------------------------------- #
# the comparison-result cache (verdict + discrepancy counts), keyed by baseline
# --------------------------------------------------------------------------- #
def load_results(dest, baseline_key):
    """{row_key: {cell_key: {verdict, diff_cells, one_sided, built_at_mtime}}}.
    Tolerant: missing/corrupt/old-version -> {} (never raises). The cache is a
    versioned envelope (cache_envelope); a pre-P1 raw dict reads as empty (a
    one-time rebuild)."""
    p = _results_path(dest, baseline_key)
    try:
        with open(p, encoding="utf-8") as f:
            data = json.load(f)
        return cache_envelope.unwrap(data, output_identity=baseline_key)
    except OSError:
        return {}                            # not written yet (first run) — expected
    except ValueError as e:                  # corrupt JSON: surface it, then degrade
        log.warning("matrix: corrupt results cache %s (%s: %s); treating as empty",
                    p, type(e).__name__, e)
        return {}


def _require_cache_guard(commit_guard, path, action):
    """Raise when a target-aware ownership guard no longer authorizes ``path``.

    Cache writes are part of the comparison publication, not disposable
    telemetry.  Silently dropping a denied write would leave the workbook and
    its rendered verdict out of sync, so ownership loss is an operation error.
    """
    if not consolidation_meta.guard_allows(commit_guard, path):
        raise ValueError(
            "The comparisons destination changed before the matrix "
            f"{action}. The cache was left untouched; retry the comparison.")


def _save_results(dest, baseline_key, data, commit_guard=None):
    p = _results_path(dest, baseline_key)
    tmp = p.with_name(p.name + ".tmp")
    _require_cache_guard(commit_guard, p.parent, "cache directory write")
    _require_cache_guard(commit_guard, p, "cache write")
    _require_cache_guard(commit_guard, tmp, "cache temporary write")
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        _require_cache_guard(commit_guard, p, "cache write")
        _require_cache_guard(commit_guard, tmp, "cache temporary write")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(cache_envelope.wrap(data, output_identity=baseline_key), f)
        _require_cache_guard(commit_guard, p, "cache publication")
        _require_cache_guard(commit_guard, tmp, "cache publication")
        os.replace(tmp, p)
    except OSError as e:
        log.warning("matrix: could not write results cache %s: %s: %s",
                    p, type(e).__name__, e)
        raise ValueError(
            "The comparison workbook was created, but its Matrix result cache "
            "could not be safely published. Refresh the cell.") from e


def record_result(dest, baseline_key, row_key, cell_key, verdict,
                  diff_cells, one_sided, built_at_mtime, completion=None,
                  input_fingerprint=None, generation_id=None,
                  producer_versions=None, commit_guard=None):
    data = load_results(dest, baseline_key)
    data.setdefault(row_key, {})[cell_key] = {
        "verdict": verdict, "diff_cells": diff_cells,
        "one_sided": one_sided, "built_at_mtime": built_at_mtime,
        "completion": completion,        # P1-R01: partial inputs flagged durably
        "generation_id": generation_id,
        # P2/F5: the cell's input-folder identity at build time; a later snapshot reads
        # the cell STALE when it differs (a route added/removed/resized — invisible to
        # the mtime check). Absent on legacy records -> the mtime check alone applies.
        "input_fingerprint": input_fingerprint,
        # CMP-AUD-084: the semantic producer version this cell was built by. A later
        # snapshot reads the cell stale when the running pipeline differs (a shipped
        # comparator/parser fix); absent on legacy records -> stale once, then rebuilt.
        "producer_versions": producer_versions,
    }
    _save_results(dest, baseline_key, data, commit_guard=commit_guard)


# --------------------------------------------------------------------------- #
# read discrepancy counts back from a produced VALUES workbook (no COM/F9)
# --------------------------------------------------------------------------- #
def read_counts(values_path, has_route=None):
    """Read ``(diff_cells, one_sided)`` from a VALUES comparison workbook.

    ``Status`` and ``Diffs`` are the structured truth columns emitted by every
    current comparison layout. Visible field text is never inspected: the
    spaced not-equal glyph is legitimate source content and cannot encode state.
    A missing, duplicate, or malformed count contract returns ``(None, None)``
    instead of guessing a layout or silently certifying zero differences.

    Label lookup remains correct whether column A is the Route-keyed layout's
    route or a flat aggregate whose identity field happens to be named Route.
    ``has_route`` is retained only for call compatibility and cannot authorize a
    positional fallback.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(values_path, read_only=True, data_only=True)
    except Exception as e:                       # noqa: BLE001 (best-effort read)
        log.debug("read_counts: can't open %s (%s: %s)", values_path,
                  type(e).__name__, e)
        return (None, None)
    try:
        ws = wb["Comparison"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None) or ()     # row 1 (header); data follows

        def _col_of(label):                      # unique 1-based exact label
            matches = [i + 1 for i, value in enumerate(header)
                       if value == label]
            return matches[0] if len(matches) == 1 else None

        status_col = _col_of("Status")
        diffs_col = _col_of("Diffs")
        if status_col is None or diffs_col is None:
            log.debug("read_counts: Comparison lacks unique Status/Diffs labels")
            return (None, None)
        one_sided = diff_cells = 0
        for row in rows_iter:                     # data rows (header consumed above)
            if row is None or all(v is None for v in row):
                continue
            status = row[status_col - 1] if len(row) >= status_col else None
            diffs = row[diffs_col - 1] if len(row) >= diffs_col else None
            if status == "Both":
                if (isinstance(diffs, bool) or not isinstance(diffs, (int, float))
                        or not float(diffs).is_integer() or diffs < 0):
                    log.debug("read_counts: matched row has invalid Diffs value %r", diffs)
                    return (None, None)
                diff_cells += int(diffs)
            elif isinstance(status, str) and status:
                if diffs not in (None, ""):
                    log.debug("read_counts: one-sided row unexpectedly carries Diffs %r", diffs)
                    return (None, None)
                one_sided += 1
            else:
                log.debug("read_counts: row has invalid Status value %r", status)
                return (None, None)
        return (diff_cells, one_sided)
    except Exception as e:                       # noqa: BLE001
        log.debug("read_counts: can't read %s (%s: %s)", values_path,
                  type(e).__name__, e)
        return (None, None)
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# per-cell comparison freshness (decision c: mtime staleness + P2 input identity)
# --------------------------------------------------------------------------- #
def _cell_input_fingerprint(*folders):
    """A combined identity over a comparison cell's TSMIS source FOLDER(s) — the
    multi-file side(s) where a DELETED route hides from a newest-mtime signal — joined
    in the given order. Build-time (record_*) and read-time (the freshness readers) MUST
    pass the same folders in the same order so the strings agree. (FILE sides — a TSN /
    baseline workbook — are captured by their mtime in the existing freshness check.)"""
    return "|".join(artifact_store.fingerprint(Path(f)) for f in folders)


def _inputs_changed(rec_trusted, rec, *folders):
    """True iff a TRUSTED cache record carries a recorded input fingerprint that differs
    from the cell's CURRENT source-folder identity (R1-R03: a route added / removed /
    resized since the comparison was built). False here when the record is untrusted or
    has no recorded fingerprint; the caller treats either condition as
    ``cache_missing_or_mismatched`` and therefore stale/rebuildable rather than falling
    back to mtime-only trust."""
    if not rec_trusted or not rec:
        return False
    rec_fp = rec.get("input_fingerprint")
    if not rec_fp:
        return False
    return _cell_input_fingerprint(*folders) != rec_fp


def _nested_record(results, outer_key, inner_key):
    """Read one nested cache record without trusting JSON container shapes."""
    outer = results.get(outer_key, {}) if isinstance(results, dict) else {}
    return outer.get(inner_key) if isinstance(outer, dict) else None


def _published_comparison_result(path, result):
    """Return the strict on-disk record for one just-built comparison.

    A successful producer result is not enough: the Matrix cache may advertise
    truth only after the workbook, typed outcome, and complete generation sidecar
    agree.  This is also the shared boundary used by validation/day/baseline.
    """
    try:
        return consolidation_meta.require_published_comparison(path, result)
    except ValueError as e:
        raise ValueError(
            "The comparison workbook was created, but its published outcome "
            f"cannot be trusted. Refresh the comparison before using it. ({e})") from e


def _staleness(cmp_m, sources, rec, fp_folders, missing_side,
               comparison_output=None):
    """The staleness verdict both cell readers share (S4 / ARC-03):
    {supported, built, mtime, stale, reason, missing_side, verdict, diff_cells,
    one_sided, completion}. `sources` is [{name, mtime}, ...] in reason order
    ("<name>_newer" / "both_newer"); the cached verdict/counts surface only
    while the record's built_at_mtime still matches the workbook; a trusted
    record's input fingerprint must match the current `fp_folders` identity
    (R1-R03). When `comparison_output` is supplied (all production snapshots),
    verdict/counts/completion come only from its strict generation sidecar; the
    cache is retained solely for source-freshness metadata. Callers own
    `missing_side` (their side semantics differ)."""
    built = cmp_m is not None
    verdict = diff_cells = one_sided = completion = pairing_quality = None
    publication_reason = None
    published_generation = None
    if built and comparison_output is not None:
        published = consolidation_meta.read_comparison_outcome(comparison_output)
        if published is None:
            publication_reason = "outcome_missing"
        elif not published.trusted:
            publication_reason = "outcome_untrusted"
            completion = outcome.PARTIAL
        else:
            typed = published.comparison_outcome
            verdict = typed.verdict
            diff_cells = typed.counts.differing_cells
            one_sided = (typed.counts.side_a_only_rows
                         + typed.counts.side_b_only_rows)
            completion = typed.completion
            pairing_quality = typed.pairing_quality
            published_generation = published.artifact_generation.generation_id
    rec_is_mapping = isinstance(rec, dict)
    try:
        rec_mtime = float(rec.get("built_at_mtime", -1)) if rec_is_mapping else None
    except (TypeError, ValueError, OverflowError):  # silent-ok: malformed cache mtime makes the record untrusted
        rec_mtime = None
    rec_trusted = bool(built and rec_is_mapping and rec_mtime is not None
                       and abs(rec_mtime - cmp_m) < _MTIME_TOL_S)
    if rec_trusted and fp_folders:
        recorded_fingerprint = rec.get("input_fingerprint")
        if not isinstance(recorded_fingerprint, str) or not recorded_fingerprint:
            rec_trusted = False
    if (rec_trusted and comparison_output is not None
            and rec.get("generation_id") != published_generation):
        rec_trusted = False
    # CMP-AUD-084: a cache stamped by a DIFFERENT semantic producer generation (a
    # shipped comparator / parser / normalizer fix) can never certify fresh, even
    # with byte-identical inputs and a matching output generation. The recorded
    # producer identity must equal the CURRENT one; a legacy record (no field, or a
    # mismatching map) reads stale and rebuilds ONCE against the running pipeline.
    producer_stale = bool(rec_trusted
                          and rec.get("producer_versions") != producer_identity())
    if producer_stale:
        rec_trusted = False
    if rec_trusted and comparison_output is None:
        verdict = rec.get("verdict")
        diff_cells = rec.get("diff_cells")
        one_sided = rec.get("one_sided")
        # P1-R01: a cell built from PARTIAL inputs carries it durably — the
        # matrix must flag "compared, but inputs were incomplete". Old records
        # (no field) read as complete.
        completion = rec.get("completion")
    recorded_ids = rec.get("source_identities", {}) if rec_trusted else {}
    if not isinstance(recorded_ids, dict):
        recorded_ids = {}
    current_ids = {s["name"]: s.get("identity") for s in sources
                   if s.get("identity") is not None}
    identity_changed = False
    if recorded_ids:
        identity_changed = recorded_ids != current_ids
    elif any(s.get("identity_required") for s in sources):
        # A pre-identity cache cannot certify that it used the selected bytes.
        identity_changed = True

    if not built:
        stale, reason = True, "missing"
    elif publication_reason is not None:
        stale, reason = True, publication_reason
    elif producer_stale:
        # CMP-AUD-084: the record was built by a superseded semantic pipeline.
        stale, reason = True, "producer_version_changed"
    elif comparison_output is not None and not rec_trusted:
        # The strict sidecar owns output truth, but the per-cell cache owns the
        # source-folder set fingerprint. Losing either half cannot certify fresh.
        stale, reason = True, "cache_missing_or_mismatched"
    elif identity_changed:
        stale, reason = True, "source_identity_changed"
    else:
        newer = [s["name"] for s in sources
                 if s.get("mtime") is not None and s["mtime"] > cmp_m + _MTIME_TOL_S]
        if newer:
            stale = True
            reason = "both_newer" if len(newer) > 1 else f"{newer[0]}_newer"
        elif _inputs_changed(rec_trusted, rec, *fp_folders):
            stale, reason = True, "inputs_changed"
        else:
            stale, reason = False, "fresh"
    if not stale and completion == outcome.PARTIAL:
        # A comparison over incomplete inputs is useful observed truth, but it
        # is never a settled green generation. Keep it visible and retryable.
        stale, reason = True, "partial"
    return {"supported": True, "built": built, "mtime": cmp_m, "stale": stale,
            "reason": reason, "missing_side": missing_side, "verdict": verdict,
            "diff_cells": diff_cells, "one_sided": one_sided,
            "completion": completion, "pairing_quality": pairing_quality}


def comparison_state(dest, baseline_key, row_key, cell_key, subdir,
                     cell_ages_map, results):
    """{built, mtime, stale, reason, missing_side, verdict, diff_cells,
    one_sided} for one non-baseline cell. STALE when the comparison file is
    missing, either side's export mtime is newer than it, OR the inputs' IDENTITY
    changed since it was built (a route added/removed/resized — F5/P2). The cached
    verdict/counts are surfaced only when the cache's recorded mtime still
    matches the file (else they read as unknown and the cell shows 're-run')."""
    dest = Path(dest)                            # accept str/Path (we join folders below)
    base_m = cell_ages_map.get(baseline_key, {}).get(subdir, {}).get("mtime")
    cell_m = cell_ages_map.get(cell_key, {}).get(subdir, {}).get("mtime")
    if base_m is None and cell_m is None:
        missing_side = "both"
    elif base_m is None:
        missing_side = "baseline"
    elif cell_m is None:
        missing_side = "cell"
    else:
        missing_side = None

    cmp_path = comparison_path(dest, baseline_key, row_key, cell_key)
    try:
        cmp_m = cmp_path.stat().st_mtime
    except OSError:
        cmp_m = None
    rec = _nested_record(results, row_key, cell_key)
    # reason order matters: [baseline, cell] -> "baseline_newer"/"cell_newer".
    return _staleness(cmp_m,
                      [{"name": "baseline", "mtime": base_m},
                       {"name": "cell", "mtime": cell_m}],
                      rec,
                      (dest / cell_key / subdir, dest / baseline_key / subdir),
                      missing_side,
                      comparison_output=cmp_path)


# --------------------------------------------------------------------------- #
# TSN / cross-format modes. The matrix only ORCHESTRATES the existing (manual)
# comparison adapters — compare_env / compare_highway_log / compare_highway_log_pdf
# are NEVER edited. TSN drops: <dest>/_tsn_input/<subdir>/; the TSN + cross-format
# comparison sheets: <dest>/comparisons/tsn/ (apart from the cross-env tree).
# --------------------------------------------------------------------------- #
TSN_INPUT_DIRNAME = "_tsn_input"
TSN_COMPARE_DIRNAME = "tsn"            # under comparisons/
_TSN_RESULTS_FILE = "_tsn_results.json"


def _safe_mtime(p):
    if not p:
        return None
    try:
        return Path(p).stat().st_mtime
    except OSError:
        return None


def tsn_input_root(dest, subdir):
    """Where the user drops the TSN dataset (a consolidated workbook or the
    district PDFs). TSN is ONE dataset per report family, so both Highway Log TSN
    modes (Excel-vs-TSN, PDF-vs-TSN) share <dest>/_tsn_input/highway_log/."""
    return Path(dest) / TSN_INPUT_DIRNAME / subdir


def tsn_comparisons_root(dest):
    return Path(dest) / COMPARISONS_DIRNAME / TSN_COMPARE_DIRNAME


def tsn_source(dest, subdir, selected_file=None):
    """Resolve the TSN dataset for a report `subdir`, returning {kind:
    file|consolidated|pdfs|raw|missing_explicit|none, path?, mtime?, pdf_count?,
    raw_count?, selected_path?}.

    Delegates to the canonical TSN library (tsn_library.resolve): an explicit
    user-picked `selected_file` wins; else the library's consolidated workbook;
    else its raw file(s) -> the 'consolidate first' state; else the legacy
    dest-scoped drop <dest>/_tsn_input/<subdir>/ and the global legacy locations
    (back-compat) — so an existing install keeps resolving until imported."""
    import tsn_library                              # lazy: no import cycle
    subdir = tsn_library.canonical_dataset_key(subdir)
    return tsn_library.resolve(subdir, selected_file, legacy_dest=dest)


# --- per-row comparison MODE registry -------------------------------------- #
# A row's "env" (cross-environment) mode is supported only when it has a
# compare_env folder adapter. Highway Log is TWO rows (Excel + PDF), each with its
# own TSN + cross-format modes; the other reports get one greyed "tsn" placeholder
# (no comparison code yet). A mode dict: id, label, kind ("env"|"tsn"|"self"),
# supported, env_subdir, plus tsn_subdir+fmt (tsn) or other_subdir (self).
def _row_modes(row_key, subdir, adapter):
    env = {"id": "env", "label": "Cross-environment", "kind": "env",
           "supported": adapter is not None, "env_subdir": subdir}
    # Every mode's `supported` DERIVES from the comparison registry (CMP-AUD-013):
    # the 'tsn' modes from tsn_supported(row_key) and the 'self' modes from
    # self_supported(<pdf subdir>) — patching the registry (tsn_comparator_for /
    # _pdf_self_comparator) now flips these rows here too, instead of a hand-written
    # True shadowing it. Production identities are unchanged (all True today).
    if row_key == "highway_log":            # TSMIS Excel
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn",
                 "supported": tsn_supported(row_key),
                 "env_subdir": "highway_log", "tsn_subdir": "highway_log", "fmt": "excel"},
                {"id": "vs_pdf", "label": "vs TSMIS PDF", "kind": "self",
                 "supported": self_supported("highway_log_pdf"),
                 "env_subdir": "highway_log", "other_subdir": "highway_log_pdf"}]
    if row_key == "highway_log_pdf":        # TSMIS PDF
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn",
                 "supported": tsn_supported(row_key),
                 "env_subdir": "highway_log_pdf", "tsn_subdir": "highway_log", "fmt": "pdf"},
                {"id": "vs_excel", "label": "vs TSMIS Excel", "kind": "self",
                 "supported": self_supported("highway_log_pdf"),
                 "env_subdir": "highway_log_pdf", "other_subdir": "highway_log"}]
    if row_key == "intersection_detail_pdf":   # the exact parallel of highway_log_pdf
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn",
                 "supported": tsn_supported(row_key),
                 "env_subdir": "intersection_detail_pdf",
                 "tsn_subdir": "intersection_detail", "fmt": "pdf"},
                {"id": "vs_excel", "label": "vs TSMIS Excel", "kind": "self",
                 "supported": self_supported("intersection_detail_pdf"),
                 "env_subdir": "intersection_detail_pdf", "other_subdir": "intersection_detail"}]
    if row_key == "highway_detail_pdf":        # the exact parallel of intersection_detail_pdf
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn",
                 "supported": tsn_supported(row_key),
                 "env_subdir": "highway_detail_pdf",
                 "tsn_subdir": "highway_detail", "fmt": "pdf"},
                {"id": "vs_excel", "label": "vs TSMIS Excel", "kind": "self",
                 "supported": self_supported("highway_detail_pdf"),
                 "env_subdir": "highway_detail_pdf", "other_subdir": "highway_detail"}]
    if row_key == "highway_sequence_pdf":      # the exact parallel of highway_detail_pdf
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn",
                 "supported": tsn_supported(row_key),
                 "env_subdir": "highway_sequence_pdf",
                 "tsn_subdir": "highway_sequence", "fmt": "pdf"},
                {"id": "vs_excel", "label": "vs TSMIS Excel", "kind": "self",
                 "supported": self_supported("highway_sequence_pdf"),
                 "env_subdir": "highway_sequence_pdf", "other_subdir": "highway_sequence"}]
    if row_key == "ramp_detail_pdf":           # the exact parallel of highway_sequence_pdf
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn",
                 "supported": tsn_supported(row_key),
                 "env_subdir": "ramp_detail_pdf",
                 "tsn_subdir": "ramp_detail", "fmt": "pdf"},
                {"id": "vs_excel", "label": "vs TSMIS Excel", "kind": "self",
                 "supported": self_supported("ramp_detail_pdf"),
                 "env_subdir": "ramp_detail_pdf", "other_subdir": "ramp_detail"}]
    return [env,
            {"id": "tsn", "label": "vs TSN", "kind": "tsn",
             "supported": tsn_supported(row_key),
             "env_subdir": subdir, "tsn_subdir": subdir, "fmt": None}]


def _mode_by_id(modes, mode_id):
    for m in modes:
        if m["id"] == mode_id:
            return m
    return modes[0]                          # default to env


def tsn_comparator_for(row_key):
    """The vs-TSN comparison adapter for a report row (a `"files"` module/instance
    exposing compare(path_a, path_b, out_path, …)), or None when the report has no
    TSN comparator yet. The single registry of which reports support vs-TSN — adding
    a report here (+ its CompareSchema module) flips it on in BOTH matrices."""
    if row_key == "highway_log":
        import compare_highway_log as _m            # lazy
        return _m
    if row_key == "highway_log_pdf":
        import compare_highway_log_pdf as _m
        return _m.TSMIS_PDF_VS_TSN
    if row_key == "intersection_detail_pdf":
        import compare_intersection_detail_pdf as _m
        return _m.TSMIS_PDF_VS_TSN
    if row_key == "ramp_detail":
        import compare_ramp_detail_tsn as _m
        return _m
    if row_key == "ramp_summary":
        import compare_ramp_summary_tsn as _m       # AGGREGATE recipe
        return _m
    if row_key == "intersection_summary":
        import compare_intersection_summary_tsn as _m   # AGGREGATE
        return _m
    if row_key == "intersection_detail":
        import compare_intersection_detail_tsn as _m    # FLAT
        return _m
    if row_key == "highway_sequence":
        import compare_highway_sequence_tsn as _m       # FLAT (county+PM key)
        return _m
    if row_key == "highway_detail":
        import compare_highway_detail_tsn as _m         # FLAT (route+canonical PM)
        return _m
    if row_key == "highway_detail_pdf":
        import compare_highway_detail_pdf as _m
        return _m.TSMIS_PDF_VS_TSN
    if row_key == "highway_sequence_pdf":
        import compare_highway_sequence_pdf as _m
        return _m.TSMIS_PDF_VS_TSN
    if row_key == "ramp_detail_pdf":
        import compare_ramp_detail_pdf as _m
        return _m.TSMIS_PDF_VS_TSN
    return None


def _pdf_self_comparator(pdf_subdir):
    """The PDF-vs-Excel self-comparison adapter for a PDF report subdir (the internal
    TSMIS PDF↔Excel consistency check). Generalizes the self mode so a second PDF
    report (Intersection Detail) resolves the same way Highway Log does, instead of a
    literal `compare_highway_log_pdf` reference."""
    if pdf_subdir == "highway_log_pdf":
        import compare_highway_log_pdf as _m
        return _m.TSMIS_PDF_VS_EXCEL
    if pdf_subdir == "intersection_detail_pdf":
        import compare_intersection_detail_pdf as _m
        return _m.TSMIS_PDF_VS_EXCEL
    if pdf_subdir == "highway_detail_pdf":
        import compare_highway_detail_pdf as _m
        return _m.TSMIS_PDF_VS_EXCEL
    if pdf_subdir == "highway_sequence_pdf":
        import compare_highway_sequence_pdf as _m
        return _m.TSMIS_PDF_VS_EXCEL
    if pdf_subdir == "ramp_detail_pdf":
        import compare_ramp_detail_pdf as _m
        return _m.TSMIS_PDF_VS_EXCEL
    raise ValueError(f"no PDF-vs-Excel comparator for {pdf_subdir}")


def self_supported(pdf_subdir):
    """True when a PDF report subdir has a coded PDF-vs-Excel self comparator
    (_pdf_self_comparator). The registry SoT for the matrix 'self' mode, the exact
    parallel of tsn_supported for the 'tsn' mode — so _row_modes DERIVES a mode's
    support from the registry instead of hardcoding it (CMP-AUD-013), and a re-
    hardcoded row that drifts from the registry is caught by the parity guard."""
    try:
        return _pdf_self_comparator(pdf_subdir) is not None
    except ValueError:  # silent-ok: the raise IS the "no self comparator" signal (parallels tsn_supported's `is not None`)
        return False


def tsn_supported(row_key):
    """True when the report row has a coded vs-TSN comparator (the matrices show it
    live). Every report qualifies as of v0.17.0; an uncoded row would grey out —
    defensive, kept for any future report added before its comparator lands."""
    return tsn_comparator_for(row_key) is not None


def tsn_subdir_for(row_key, subdir, adapter=None):
    """The TSN dataset key (per-row `tsn_subdir`) for a row's vs-TSN comparison.
    Both Highway Log rows share ONE TSN dataset ('highway_log'); every other report
    uses its own subdir. The single source of truth is the row's 'tsn' mode in
    _row_modes — this is what replaces day_matrix's hardcoded TSN_SUBDIR."""
    for m in _row_modes(row_key, subdir, adapter):
        if m.get("kind") == "tsn":
            return m.get("tsn_subdir") or subdir
    return subdir


def tsn_capable(row_key):
    """True if the row offers a coded comparison beyond cross-environment (every
    report as of v0.17.0; originally only the two Highway Log rows). Drives the
    'tsn_capable' chip hint + the 'set all to vs TSN' bulk action."""
    defs = _row_defs()
    if row_key not in defs:
        return False
    _l, subdir, _i, adapter, _hr = defs[row_key]
    return any(m["kind"] != "env" and m["supported"]
               for m in _row_modes(row_key, subdir, adapter))


# --- comparison-output paths + the non-env (TSN/self) counts cache --------- #
def mode_out_path(dest, baseline_key, row_key, cell_key, mode):
    """Where a cell's comparison workbook lives. Cross-env stays under
    comparisons/<baseline>/ (existing); TSN/self modes go to comparisons/tsn/ with
    the mode in the name so flavors coexist."""
    if mode["kind"] == "env":
        return comparison_path(dest, baseline_key, row_key, cell_key)
    return tsn_comparisons_root(dest) / f"{cell_key}_{row_key}_{mode['id']}.xlsx"


def out_path_for_cell(dest, baseline_key, row_key, cell_key, mode_id, row_defs=None):
    """The comparison workbook path for one cell under the row's mode — for the
    'open' action. None for an unknown row."""
    rows = row_defs if row_defs is not None else _row_defs()
    if row_key not in rows:
        return None
    _l, subdir, _i, adapter, _hr = rows[row_key]
    mode = _mode_by_id(_row_modes(row_key, subdir, adapter), mode_id)
    return mode_out_path(dest, baseline_key, row_key, cell_key, mode)


def _tsn_results_path(dest):
    return tsn_comparisons_root(dest) / _TSN_RESULTS_FILE


def load_tsn_results(dest):
    """{ "<row>|<mode>": {cell_key: {verdict, diff_cells, one_sided,
    built_at_mtime}} } — the non-env (TSN/self) comparison counts cache."""
    p = _tsn_results_path(dest)
    try:
        with open(p, encoding="utf-8") as f:
            data = json.load(f)
        return cache_envelope.unwrap(data, output_identity="tsn")
    except OSError:
        return {}                            # not written yet (first run) — expected
    except ValueError as e:                  # corrupt JSON: surface it, then degrade
        log.warning("matrix: corrupt TSN results cache %s (%s: %s); treating as empty",
                    p, type(e).__name__, e)
        return {}


def record_tsn_result(dest, result_key, cell_key, verdict, diff_cells, one_sided,
                      built_at_mtime, completion=None, input_fingerprint=None,
                      source_identities=None, generation_id=None,
                      producer_versions=None, commit_guard=None):
    data = load_tsn_results(dest)
    data.setdefault(result_key, {})[cell_key] = {
        "verdict": verdict, "diff_cells": diff_cells,
        "one_sided": one_sided, "built_at_mtime": built_at_mtime,
        "completion": completion,        # P1-R01: partial inputs flagged durably
        "generation_id": generation_id,
        # P2/F5: the cell's TSMIS source-folder identity at build time; a later snapshot
        # reads the cell stale when it differs. Absent on legacy records (mtime only).
        "input_fingerprint": input_fingerprint,
        "source_identities": source_identities or {},
        # CMP-AUD-084: the semantic producer version (see record_result) — a shipped
        # comparator/parser change reads the cell stale even with identical inputs.
        "producer_versions": producer_versions,
    }
    p = _tsn_results_path(dest)
    tmp = p.with_name(p.name + ".tmp")
    _require_cache_guard(commit_guard, p.parent, "TSN cache directory write")
    _require_cache_guard(commit_guard, p, "TSN cache write")
    _require_cache_guard(commit_guard, tmp, "TSN cache temporary write")
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        _require_cache_guard(commit_guard, p, "TSN cache write")
        _require_cache_guard(commit_guard, tmp, "TSN cache temporary write")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(cache_envelope.wrap(data, output_identity="tsn"), f)
        _require_cache_guard(commit_guard, p, "TSN cache publication")
        _require_cache_guard(commit_guard, tmp, "TSN cache publication")
        os.replace(tmp, p)
    except OSError as e:
        log.warning("matrix: could not write TSN results cache %s: %s: %s",
                    p, type(e).__name__, e)
        raise ValueError(
            "The comparison workbook was created, but its TSN Matrix result cache "
            "could not be safely published. Refresh the cell.") from e


# --- unified per-cell comparison state ------------------------------------- #
def _cmp_state(out_path, sources, rec, fp_folders=()):
    """{supported, built, mtime, stale, reason, missing_side, verdict, diff_cells,
    one_sided} for one comparison cell. `sources` is the input sides
    [{name, present, mtime}]; STALE when the workbook is missing, any present source is
    newer than it, OR the inputs' IDENTITY changed (`fp_folders` — the TSMIS store
    folder(s); a route added/removed/resized that the mtime check misses — F5/P2)."""
    missing = [s["name"] for s in sources if not s.get("present")]
    missing_side = missing[0] if missing else None
    return _staleness(_safe_mtime(out_path), sources, rec, fp_folders, missing_side,
                      comparison_output=out_path)


# --------------------------------------------------------------------------- #
# the snapshot the GUI renders (pure filesystem read)
# --------------------------------------------------------------------------- #
def apply_order(keys, order):
    """Reorder `keys` by the user's preference list `order`: keys named in `order`
    come first (in that order, only those actually present), then any remaining keys
    in their original order. `order` is a PREFERENCE, not a filter — unknown/stale
    keys in `order` are ignored, and keys missing from `order` are kept (appended in
    their natural order), so a report/env added or removed later degrades gracefully."""
    if not order:
        return list(keys)
    keyset = set(keys)
    front = [k for k in order if k in keyset]
    frontset = set(front)
    return front + [k for k in keys if k not in frontset]


def matrix_snapshot(dest, baseline_key=BASELINE_DEFAULT, envs=None,
                    env_labels=None, row_defs=None, hidden=None, hidden_envs=None,
                    row_modes=None, tsn_files=None, now=None,
                    row_order=None, env_order=None):
    """The full render model for the matrix. PURE stat — no workbook opened (counts
    come from the caches). Per row, the SELECTED mode (`row_modes`, default 'env')
    decides each cell's comparison; `hidden`/`hidden_envs` drop rows/columns from
    the rendered+refreshed grid (still listed in all_rows/all_envs for the
    toggles); `row_order`/`env_order` are the user's drag-to-reorder preferences
    applied to the VISIBLE rows/columns. `row_defs` is injectable for tests."""
    dest = Path(dest)                            # accept str/Path (we join store folders below)
    now = now if now is not None else time.time()
    all_defs = row_defs if row_defs is not None else _row_defs()
    hidden = set(hidden or [])
    hidden_envs = set(hidden_envs or [])
    row_modes = row_modes or {}
    import tsn_library                              # lazy: canonical selection view
    tsn_files, _selection_map_changed = tsn_library.canonicalize_selections(
        tsn_files or {})
    rows = {k: v for k, v in all_defs.items() if k not in hidden}
    rows = {k: rows[k] for k in apply_order(list(rows.keys()), row_order)}
    all_envs = list(envs) if envs is not None else env_keys()
    envs = apply_order([e for e in all_envs if e not in hidden_envs], env_order)
    env_labels = env_labels or {k: default_env_label(k) for k in all_envs}

    # Resolve each visible row's selected mode + gather the store subdirs whose
    # freshness the cells need (the mode's env side + any 'other' side).
    sel, needed = {}, set()
    for row_key, (label, subdir, _idx, adapter, _hr) in rows.items():
        mode = _mode_by_id(_row_modes(row_key, subdir, adapter),
                           row_modes.get(row_key, "env"))
        sel[row_key] = mode
        needed.add(mode.get("env_subdir", subdir))
        if mode.get("other_subdir"):
            needed.add(mode["other_subdir"])
    ages = report_library.cell_ages(dest, [(sd, sd) for sd in needed], envs, now=now)
    results = load_results(dest, baseline_key)
    tsn_results = load_tsn_results(dest)
    _absent = {"present": False, "mtime": None, "age_seconds": None}

    cells, modes_sel, modes_avail, tsn_meta = {}, {}, {}, {}
    tsn_cache = {}
    for row_key, (label, subdir, _idx, adapter, _hr) in rows.items():
        mode = sel[row_key]
        modes_sel[row_key] = mode["id"]
        modes_avail[row_key] = [{"id": m["id"], "label": m["label"],
                                 "kind": m["kind"], "supported": m["supported"]}
                                for m in _row_modes(row_key, subdir, adapter)]
        env_subdir = mode.get("env_subdir", subdir)
        src = None
        if mode["kind"] == "tsn":
            tsn_key = tsn_library.canonical_dataset_key(mode["tsn_subdir"])
            if tsn_key not in tsn_cache:
                tsn_cache[tsn_key] = (tsn_source(dest, tsn_key, tsn_files.get(tsn_key))
                                      if mode["supported"] else {"kind": "none"})
            src = tsn_cache[tsn_key]
            selected_path = tsn_library.selection_path(tsn_files.get(tsn_key))
            tsn_meta[row_key] = {"supported": mode["supported"], "fmt": mode.get("fmt"),
                                 "source_kind": src.get("kind"), "source_path": src.get("path"),
                                 "pdf_count": src.get("pdf_count"),
                                 "tsn_subdir": tsn_key,
                                 "file": selected_path,
                                 "selected_path": src.get("selected_path"),
                                 "selection_missing": src.get("kind") == "missing_explicit",
                                 "selection_reason": src.get("selection_reason"),
                                 "input_dir": str(tsn_input_root(dest, tsn_key))}
        per = {}
        for env in envs:
            export = ages.get(env, {}).get(env_subdir, _absent)
            is_baseline = (mode["kind"] == "env" and env == baseline_key)
            if not mode["supported"]:
                cmp = {"supported": False}
            elif is_baseline:
                cmp = None
            elif mode["kind"] == "env":
                cmp = comparison_state(dest, baseline_key, row_key, env, env_subdir,
                                       ages, results)
            elif mode["kind"] == "tsn":
                rec = _nested_record(tsn_results, f"{row_key}|{mode['id']}", env)
                sources = [{"name": "cell", "present": export["present"], "mtime": export["mtime"]},
                           {"name": "tsn",
                            "present": src.get("kind") in ("file", "consolidated"),
                            "mtime": src.get("mtime"),
                            "identity": src.get("identity_token"),
                            "identity_required": True}]
                # F5/P2: the TSMIS store folder is the multi-file side where a deleted
                # route hides from mtime — fingerprint it (the TSN side is a file, mtime).
                cmp = _cmp_state(mode_out_path(dest, baseline_key, row_key, env, mode),
                                 sources, rec, fp_folders=(dest / env / env_subdir,))
            else:                            # self: TSMIS PDF vs Excel
                other = ages.get(env, {}).get(mode["other_subdir"], _absent)
                rec = _nested_record(tsn_results, f"{row_key}|{mode['id']}", env)
                sources = [{"name": "cell", "present": export["present"], "mtime": export["mtime"]},
                           {"name": "other", "present": other["present"], "mtime": other["mtime"]}]
                # F5/P2: both sides are TSMIS store folders — fingerprint both (env first,
                # other second; build_comparison records them in this same order).
                cmp = _cmp_state(mode_out_path(dest, baseline_key, row_key, env, mode),
                                 sources, rec,
                                 fp_folders=(dest / env / env_subdir,
                                             dest / env / mode["other_subdir"]))
            cell = {"export": export, "is_baseline": is_baseline, "cmp": cmp}
            if mode["kind"] == "env":
                cell["comparison"] = cmp     # back-compat alias for the Stage-A UI
            per[env] = cell
        cells[row_key] = per

    return {
        "dest": str(dest),
        "baseline": baseline_key,
        "rows": list(rows.keys()),
        "row_labels": {k: rows[k][0] for k in rows},
        # Every matrix row (visible or not) for the toggle chips; `tsn_capable`
        # marks the rows with a coded comparison beyond cross-env.
        "all_rows": [{"key": k, "label": all_defs[k][0],
                      "tsn_capable": tsn_capable(k)} for k in all_defs],
        "hidden": sorted(hidden),
        "modes": modes_sel,            # row_key -> selected mode id
        "row_modes": modes_avail,      # row_key -> [{id,label,kind,supported}]
        "tsn_meta": tsn_meta,          # row_key -> TSN source summary (tsn mode only)
        "envs": list(envs),
        "all_envs": all_envs,
        "hidden_envs": sorted(hidden_envs),
        "env_labels": env_labels,
        "cells": cells,
    }
