"""The ArcGIS per-layer library substrate (Clean Road, v0.29.0).

Reads the owner's per-layer exports in `arcgis_layers/` — `NN_<Layer Name>.xlsx`
files beside a `00_INDEX.xlsx` manifest — and gives the clean-road consolidators
the shared machinery they all need:

  * the INDEX manifest (identity + audit provenance: each layer's FeatureServer
    `Data Source`), with a row-count sanity gate — a file with FEWER data rows
    than its own INDEX claims is a truncated export and refuses; a few MORE is a
    measured ArcGIS count/export race on healthy exports (the 2026-07-22 drop:
    Travel Way L +1, County Code +3, City +1) and passes with the delta recorded;
  * name-keyed streaming of any layer's columns (headers vary per layer, so
    position is never trusted);
  * the two export DIALECTS, normalized to the TSN/TASAS codes: the per-layer
    exports resolve coded domains to LABELS ("District 7" / "Los Angeles" /
    "Right") where the earlier bundle exports carried CODES ("7" / "LA." / "R");
  * LRS time algebra — a slice is current iff `LRSToDate` is empty, and the
    as-of-D selection is `LRSFromDate <= D < LRSToDate` (the owner's rule,
    value-proven on route 001);
  * postmile interval algebra in EXACT integer micro-postmiles (1e-5 mi — THY
    carries up to 5 decimals), so breakpoint unions never suffer float fuzz.

Console-free: raises ValueError / returns values, never prints. openpyxl is
imported lazily so the module loads in bare environments.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import paths

INDEX_NAME = "00_INDEX.xlsx"
INDEX_HEADER = ["Excel File", "ArcGIS Layer or Table", "Rows Exported",
                "Fields Exported", "ArcGIS Contents Path", "Data Source"]

# The agreed 40-layer library manifest (docs/planning/cleanroad-highways.md),
# tagged by which clean-road file(s) read each layer: H(ighways),
# I(ntersections), R(amps). The GUI status card and the consolidators' missing-
# layer messages key off this — an .xlsx matching no entry is UNKNOWN (the
# drop-zone must not accumulate dead weight).
EXPECTED_LAYERS = {
    "City": "HIR",
    "County Code": "H",
    "Equation Points": "H",
    "IM Complex Intersection Cross Reference": "I",
    "IM Complex Intersection Influence Segments": "I",
    "IM Intersection Approach Detail": "I",
    "IM Intersection Approach Segments": "I",
    "IM Intersection Detail": "I",
    "IM Intersection Point": "I",
    "IM Intersection Route Table": "I",
    "Route Direction": "R",
    "SHS Access Control": "H",
    "SHS Barrier": "H",
    "SHS Curb Landscape": "H",
    "SHS Design Speed": "H",
    "SHS District": "H",
    "SHS Forest HWY": "H",
    "SHS Highway Group": "HIR",
    "SHS I Shld Width L": "H",
    "SHS I Shld Width R": "H",
    "SHS Inv Network Date": "HIR",
    "SHS Landmark": "H",
    "SHS Median": "H",
    "SHS Non Add Mileage": "H",
    "SHS O Shld Width L": "H",
    "SHS O Shld Width R": "H",
    "SHS Population": "HIR",
    "SHS Ramp": "R",
    "SHS Ramp Pt": "R",
    "SHS Route Break": "H",
    "SHS Special Feature L": "H",
    "SHS Special Feature R": "H",
    "SHS Surface Type L": "H",
    "SHS Surface Type R": "H",
    "SHS Tolls": "H",
    "SHS Travel Way L": "H",
    "SHS Travel Way R": "H",
    "Terrain Type": "H",
    "Traffic Volume Ramps": "R",
    "Traffic Volume Segments": "HI",
}

_NN_RE = re.compile(r"^(?:\d+_)?(.+)\.xlsx$", re.IGNORECASE)
# A truncated export refuses; a small over-count is the measured healthy race.
_OVERCOUNT_SLACK = 5


def root():
    return paths.ARCGIS_LAYERS_ROOT


def layer_name_of(filename):
    """The layer identity carried by a library filename (`NN_` prefix optional;
    the FILENAME is the identity — worksheet names truncate at 31 chars)."""
    m = _NN_RE.match(Path(filename).name)
    return m.group(1) if m else None


def inventory(lib_root=None):
    """The library's current stock vs the agreed manifest: `{present: {layer:
    Path}, missing: [layer], unknown: [filename], index: Path|None}`. Pure
    filesystem; never raises on an unreadable folder (empty inventory)."""
    lib = Path(lib_root or root())
    present, unknown, index_path = {}, [], None
    try:
        entries = sorted(p for p in lib.glob("*.xlsx") if p.is_file())
    except OSError:      # silent-ok: an unreadable library reads as empty; the status card then says what's missing
        entries = []
    for p in entries:
        if p.name.lower() == INDEX_NAME.lower():
            index_path = p
            continue
        name = layer_name_of(p.name)
        if name in EXPECTED_LAYERS and name not in present:
            present[name] = p
        else:
            unknown.append(p.name)
    missing = [n for n in EXPECTED_LAYERS if n not in present]
    return {"present": present, "missing": missing, "unknown": unknown,
            "index": index_path}


def read_index(lib_root=None):
    """The `00_INDEX.xlsx` manifest rows: `{layer: {file, rows, fields, path,
    source}}`. The INDEX is the audit-provenance record (`source` = the
    FeatureServer URL + layer id) and the row-count reference. Raises ValueError
    when absent or malformed — the library is verified, never guessed."""
    from openpyxl import load_workbook

    lib = Path(lib_root or root())
    p = lib / INDEX_NAME
    if not p.is_file():
        raise ValueError(
            f"The ArcGIS layer library has no {INDEX_NAME} manifest in:\n{lib}\n"
            "Export the layers with their INDEX (the per-layer export writes it) "
            "and copy all of the files in together.")
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        it = wb.worksheets[0].iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it, ())]
        if header[:len(INDEX_HEADER)] != INDEX_HEADER:
            raise ValueError(
                f"{INDEX_NAME} does not carry the expected manifest header "
                f"{INDEX_HEADER} — re-export the layer library.")
        out = {}
        for r in it:
            if not r or r[0] is None or str(r[0]).strip() == "":
                continue
            fname, layer = str(r[0]).strip(), str(r[1]).strip()
            try:
                rows, fields = int(r[2]), int(r[3])
            except (TypeError, ValueError):
                raise ValueError(f"{INDEX_NAME}: the row/field counts for "
                                 f"{fname!r} are not numbers — re-export the "
                                 "layer library.")
            out[layer] = {"file": fname, "rows": rows, "fields": fields,
                          "path": str(r[4] or ""), "source": str(r[5] or "")}
        if not out:
            raise ValueError(f"{INDEX_NAME} lists no layers — re-export the "
                             "layer library.")
        return out
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# dialect normalization (labels export <-> bundle codes <-> TSN/TASAS codes)
# --------------------------------------------------------------------------- #
COUNTY_CODES = {
    "ALAMEDA": "ALA", "ALPINE": "ALP", "AMADOR": "AMA", "BUTTE": "BUT",
    "CALAVERAS": "CAL", "COLUSA": "COL", "CONTRA COSTA": "CC",
    "DEL NORTE": "DN", "EL DORADO": "ED", "FRESNO": "FRE", "GLENN": "GLE",
    "HUMBOLDT": "HUM", "IMPERIAL": "IMP", "INYO": "INY", "KERN": "KER",
    "KINGS": "KIN", "LAKE": "LAK", "LASSEN": "LAS", "LOS ANGELES": "LA",
    "MADERA": "MAD", "MARIN": "MRN", "MARIPOSA": "MPA", "MENDOCINO": "MEN",
    "MERCED": "MER", "MODOC": "MOD", "MONO": "MNO", "MONTEREY": "MON",
    "NAPA": "NAP", "NEVADA": "NEV", "ORANGE": "ORA", "PLACER": "PLA",
    "PLUMAS": "PLU", "RIVERSIDE": "RIV", "SACRAMENTO": "SAC",
    "SAN BENITO": "SBT", "SAN BERNARDINO": "SBD", "SAN DIEGO": "SD",
    "SAN FRANCISCO": "SF", "SAN JOAQUIN": "SJ", "SAN LUIS OBISPO": "SLO",
    "SAN MATEO": "SM", "SANTA BARBARA": "SB", "SANTA CLARA": "SCL",
    "SANTA CRUZ": "SCR", "SHASTA": "SHA", "SIERRA": "SIE", "SISKIYOU": "SIS",
    "SOLANO": "SOL", "SONOMA": "SON", "STANISLAUS": "STA", "SUTTER": "SUT",
    "TEHAMA": "TEH", "TRINITY": "TRI", "TULARE": "TUL", "TUOLUMNE": "TUO",
    "VENTURA": "VEN", "YOLO": "YOL", "YUBA": "YUB",
}
_VALID_COUNTIES = frozenset(COUNTY_CODES.values())


def norm_county(v):
    """County to the TASAS code: full names ("Los Angeles"), dot-padded bundle
    codes ("LA.") and bare codes ("LA") all normalize to "LA". An unrecognized
    value comes back stripped/upper-cased so it surfaces rather than vanishes."""
    s = ("" if v is None else str(v)).strip().upper().rstrip(".")
    if not s:
        return ""
    if s in _VALID_COUNTIES:
        return s
    return COUNTY_CODES.get(s, s)


_DISTRICT_RE = re.compile(r"^\s*(?:DISTRICT\s+)?0*(\d{1,2})\s*$", re.IGNORECASE)


def norm_district(v):
    """District to the TSN 2-digit code: "District 7" / "7" / 7 / "07" -> "07"."""
    if v is None:
        return ""
    m = _DISTRICT_RE.match(str(v))
    if not m:
        return str(v).strip()
    return f"{int(m.group(1)):02d}"


def norm_alignment(v):
    """Alignment to R/L: "Right"/"R" -> "R", "Left"/"L" -> "L", none-ish -> ""."""
    s = ("" if v is None else str(v)).strip().upper()
    if s in ("", ".", "NONE"):
        return ""
    if s.startswith("R"):
        return "R"
    if s.startswith("L"):
        return "L"
    return s


def norm_route(v):
    """Route number to the TSN 3-digit form ("1" / 1 / "001" -> "001")."""
    s = ("" if v is None else str(v)).strip()
    if s.endswith(".0"):                      # a float-typed export cell
        s = s[:-2]
    return f"{int(s):03d}" if s.isdigit() else s.upper()


def dot_none(v):
    """The exports' '.' = none sentinel (PM prefix/suffix, route suffix)."""
    s = ("" if v is None else str(v)).strip()
    return "" if s == "." else s


_CODED_RE = re.compile(r"^\s*([A-Z0-9]{1,2})\s*-\s?")


def code_of(v):
    """A coded-domain cell to its TASAS code: the labels dialect writes
    'J- Unpaved Median' / '7- No Curbs or Shrubs' / 'N - Non-Add'; the code is
    the leading 1–2 char token before the dash. A bare code ('Z') or an
    unlabelled value passes through stripped, so nothing is invented."""
    s = ("" if v is None else str(v)).strip()
    m = _CODED_RE.match(s)
    return m.group(1) if m else s


_EXCEL_EPOCH = datetime(1899, 12, 30)
_MDY_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})")
_ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")


def to_serial(v):
    """A date-ish cell (Excel serial number, datetime/date, ISO or M/D/YYYY
    text) to the Excel serial DAY number, or None for blank/'.'/unparseable.
    LRS algebra runs on serials — the exports themselves are serial-based."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, datetime):
        return float((v - _EXCEL_EPOCH).days) + (
            v.hour * 3600 + v.minute * 60 + v.second) / 86400.0
    if isinstance(v, date):
        return float((datetime(v.year, v.month, v.day) - _EXCEL_EPOCH).days)
    s = str(v).strip()
    if s in ("", "."):
        return None
    try:
        return float(s)
    except ValueError:      # silent-ok: not a bare number — try the date shapes next
        pass
    m = _ISO_RE.match(s) or None
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = _MDY_RE.match(s)
        if not m:
            return None
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return float((datetime(y, mo, d) - _EXCEL_EPOCH).days)
    except ValueError:      # silent-ok: an impossible calendar date has no serial; the slice is treated as undated
        return None


def serial_to_date(serial):
    """An Excel serial day back to a datetime.date (None-safe)."""
    if serial is None:
        return None
    from datetime import timedelta
    return (_EXCEL_EPOCH + timedelta(days=float(serial))).date()


def is_asof(from_serial, to_serial, asof_serial):
    """The owner's LRS rule: live at D iff LRSFromDate <= D < LRSToDate, with an
    empty LRSToDate meaning current."""
    if from_serial is None or from_serial > asof_serial:
        return False
    return to_serial is None or asof_serial < to_serial


# --------------------------------------------------------------------------- #
# postmile micro-units (1e-5 mi) — exact breakpoint identity
# --------------------------------------------------------------------------- #
PM_SCALE = 100000


def pm_units(v):
    """A postmile cell to integer micro-postmiles (round-half-away at 1e-5, the
    finest precision THY carries), or None for blank/non-numeric."""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return int(round(f * PM_SCALE))


def pm_text(units):
    """Micro-postmile units back to canonical decimal text ('82.011', '0',
    '186.23798') — trailing zeros trimmed, matching compare_tsn_common
    .decimal_pm's canon."""
    if units is None:
        return ""
    neg = units < 0
    units = abs(units)
    whole, frac = divmod(units, PM_SCALE)
    text = f"{whole}.{frac:05d}".rstrip("0").rstrip(".")
    return f"-{text}" if neg and text != "0" else text


def pm_float(units):
    return None if units is None else units / PM_SCALE


# --------------------------------------------------------------------------- #
# layer streaming (name-keyed; headers vary per layer)
# --------------------------------------------------------------------------- #
def stream_layer(path, want, *, layer_name=None, expected_rows=None,
                 optional=()):
    """Yield `{column: value}` dicts for every data row of a per-layer export,
    reading ONLY the `want` columns (by NAME — layer headers differ in order).
    Columns in `optional` may be absent (they yield None — a few layers carry
    their own date columns instead of the common InventoryItemStartDate).
    Verifies the data-row count against `expected_rows` (the INDEX claim) after
    the stream completes: fewer rows than claimed is a truncated export
    (ValueError); a handful more is the measured healthy ArcGIS count race."""
    from openpyxl import load_workbook

    name = layer_name or layer_name_of(path) or Path(path).name
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else ""
                  for c in (next(it, ()) or ())]
        missing = [c for c in want if c not in header and c not in optional]
        if missing:
            raise ValueError(
                f"The {name} layer export is missing expected column(s) "
                f"{', '.join(repr(c) for c in missing)} — re-export the layer "
                "with all fields.")
        idx = {c: header.index(c) for c in want if c in header}
        absent = [c for c in want if c not in header]
        n = 0
        for row in it:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            n += 1
            rec = {c: (row[i] if i < len(row) else None)
                   for c, i in idx.items()}
            for c in absent:
                rec[c] = None
            yield rec
        if expected_rows is not None:
            if n < expected_rows:
                raise ValueError(
                    f"The {name} layer export holds {n:,} data rows but its "
                    f"INDEX claims {expected_rows:,} — the export looks "
                    "truncated; re-export the layer.")
            if n > expected_rows + max(_OVERCOUNT_SLACK, expected_rows // 1000):
                raise ValueError(
                    f"The {name} layer export holds {n:,} data rows but its "
                    f"INDEX claims {expected_rows:,} — the file does not match "
                    "its manifest; re-export the layer library together.")
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# span selection + the county+PM overlay
# --------------------------------------------------------------------------- #
def overlay(spans, cuts=()):
    """One (route, county, prefix, alignment) bucket's homogeneous segments.

    `spans` = {tag: [(begin_units, end_units, payload, rank)]} — each tag is one
    layer (or one layer column family); `rank` orders competing slices (higher
    wins; the callers pass (LRSFromDate, InventoryItemStartDate)). `cuts` adds
    extra breakpoints (landmark / equation / route-break points — TSN breaks a
    row wherever one falls). Returns `[(begin, end, {tag: payload})]` over the
    UNION of all begin/end breakpoints, each segment carrying, per tag, the
    covering span's payload (the highest-rank one when several cover —
    overlapping as-of slices exist in the wild, e.g. Traffic Volume vintages).
    """
    edge_set = set(cuts)
    for entries in spans.values():
        for b, e, _payload, _rank in entries:
            edge_set.add(b)
            edge_set.add(e)
    edges = sorted(edge_set)
    segments = []
    for b, e in zip(edges, edges[1:]):
        seg = {}
        for tag, entries in spans.items():
            best = None
            for sb, se, payload, rank in entries:
                if sb <= b and e <= se:
                    if best is None or rank > best[0]:
                        best = (rank, payload)
            if best is not None:
                seg[tag] = best[1]
        if seg:
            segments.append((b, e, seg))
    return segments


def merge_adjacent(segments, same):
    """Fuse consecutive overlay segments [(b, e, seg)] whose payloads `same()`
    judges equal AND that touch exactly (e == next b) — the overlay cuts at
    EVERY layer's breakpoints, but a row boundary exists only where a VALUE
    changes."""
    out = []
    for b, e, seg in segments:
        if out and out[-1][1] == b and same(out[-1][2], seg):
            out[-1] = (out[-1][0], e, out[-1][2])
        else:
            out.append((b, e, seg))
    return out
