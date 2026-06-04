"""Consolidate Highway Sequence Listing XLSX files into a single workbook.

Reads every XLSX in   output/highway_sequence/
Writes one workbook in output/consolidated/highway_sequence_consolidated.xlsx
with a leading "Route" column added so rows from different routes are
distinguishable in the combined file.

This script is self-contained on purpose: each consolidator owns its
own read/write logic so a quirk in one report's XLSX layout cannot
affect the others.

Importable (Phase 3b): consolidate(events, confirm_overwrite) returns a
ConsolidateResult and never prints/prompts/exits, so the GUI can drive it.
The console UX lives in cli.run_consolidate_cli, used by the __main__ entry
(and therefore by "4. consolidate (combine reports).bat").
"""
import re

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from paths import OUTPUT_ROOT
from events import Events, ConsolidateResult

INPUT_DIR = OUTPUT_ROOT / "highway_sequence"
OUT_DIR = OUTPUT_ROOT / "consolidated"
OUT_PATH = OUT_DIR / "highway_sequence_consolidated.xlsx"

# Sheet name produced by the TSMIS export — must match exactly.
SHEET_NAME = "Highway Locations"

# Pull the route token out of "highway_sequence_route_<ROUTE>.xlsx".
ROUTE_FROM_NAME = re.compile(r"_route_(\w+)\.xlsx$", re.IGNORECASE)


def extract_route(path):
    m = ROUTE_FROM_NAME.search(path.name)
    return m.group(1).upper() if m else path.stem


def read_header(path):
    """Return row 1 of the expected sheet, or None if the file isn't shaped
    like a Highway Sequence export.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            return None
        ws = wb[SHEET_NAME]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return list(row) if row else None
    finally:
        wb.close()


def stream_data_rows(path):
    """Yield each data row (row 2 onwards) as a tuple."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield row
    finally:
        wb.close()


def consolidate(events=None, confirm_overwrite=None):
    """Combine every per-route Highway Sequence XLSX into one workbook.

    Console-free: reports progress via events.on_log, asks before overwriting
    through the confirm_overwrite(path)->bool callback, and returns a
    ConsolidateResult. Honors events.is_cancelled() between files.
    """
    events = events or Events()
    if not _DEPS_OK:
        return ConsolidateResult(
            status="error",
            message='openpyxl is not installed. Run "1. setup (one time).bat" first.',
        )
    confirm = confirm_overwrite or (lambda _p: True)

    if not INPUT_DIR.exists():
        return ConsolidateResult(
            status="error",
            message=(f"Input folder is missing: {INPUT_DIR}\n"
                     'Run "3. run_export (main script).bat" and pick option 3 first.'),
        )

    files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not files:
        return ConsolidateResult(
            status="error",
            message=(f"No XLSX files found in {INPUT_DIR}\n"
                     'Run "3. run_export (main script).bat" and pick option 3 first.'),
        )

    # Confirm overwrite before reading any inputs.
    if OUT_PATH.exists() and not confirm(OUT_PATH):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"Highway Sequence Consolidation - {len(files)} file(s)")
    events.on_log("=" * 60)
    events.on_log("")

    # Lock in the first readable file's header as the canonical layout.
    # Files that disagree are skipped so misaligned columns can't silently
    # corrupt the combined workbook.
    canonical_header = None
    canonical_source = None
    for p in files:
        try:
            h = read_header(p)
        except Exception as e:
            events.on_log(f"  {p.name}: header read FAILED ({type(e).__name__}); skipping")
            continue
        if h is None:
            events.on_log(f"  {p.name}: sheet '{SHEET_NAME}' missing; skipping")
            continue
        canonical_header = h
        canonical_source = p.name
        break

    if canonical_header is None:
        return ConsolidateResult(status="error",
                                 message=f"No readable XLSX files in {INPUT_DIR}.")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Style objects are built here (not at module scope) so importing this
    # module never touches openpyxl -- the GUI can import it even if a dep is
    # missing and get a clean error result instead of an ImportError.
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # write_only mode streams rows to disk so the consolidated file can
    # be hundreds of thousands of rows without exhausting memory.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(SHEET_NAME)
    ws.freeze_panes = "B2"

    ws.column_dimensions["A"].width = 9
    for i, _ in enumerate(canonical_header, start=2):
        ws.column_dimensions[get_column_letter(i)].width = 16

    header_values = ["Route"] + [v if v is not None else "" for v in canonical_header]
    header_cells = []
    for v in header_values:
        cell = WriteOnlyCell(ws, value=str(v))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        header_cells.append(cell)
    ws.append(header_cells)

    used_files = 0
    appended_rows = 0
    skipped = []
    failed = []

    for i, p in enumerate(files, 1):
        if events.is_cancelled():
            ws.close()  # finalize the write_only lazy writer (closes its temp file)
            wb.close()  # then release the archive
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i:>3}/{len(files)}] {p.name}"
        try:
            h = read_header(p)
        except Exception as e:
            events.on_log(f"{prefix} FAILED reading header ({type(e).__name__}): {e}")
            failed.append(p.name)
            continue
        if h is None:
            events.on_log(f"{prefix} skipped: sheet '{SHEET_NAME}' missing")
            skipped.append(p.name)
            continue
        if h != canonical_header:
            events.on_log(f"{prefix} skipped: header differs from {canonical_source}")
            skipped.append(p.name)
            continue

        route = extract_route(p)
        try:
            count = 0
            for row in stream_data_rows(p):
                ws.append([route] + list(row))
                count += 1
            events.on_log(f"{prefix} +{count} rows  (route {route})")
            appended_rows += count
            used_files += 1
        except Exception as e:
            events.on_log(f"{prefix} FAILED reading rows ({type(e).__name__}): {e}")
            failed.append(p.name)

    events.on_log("")
    events.on_log("Writing consolidated workbook...")
    try:
        wb.save(OUT_PATH)
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not write {OUT_PATH.name}.\n\n"
                     "This usually means the file is open in Excel. Close it there\n"
                     "and run this consolidator again. (None of the input XLSX files\n"
                     "were modified.)"),
        )

    return ConsolidateResult(
        status="ok",
        output_path=str(OUT_PATH),
        summary_lines=[
            f"Files combined: {used_files}",
            f"Rows added:     {appended_rows}",
            f"Files skipped:  {len(skipped)} {skipped if skipped else ''}",
            f"Files failed:   {len(failed)} {failed if failed else ''}",
            f"Output file:    {OUT_PATH}",
        ],
    )


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
