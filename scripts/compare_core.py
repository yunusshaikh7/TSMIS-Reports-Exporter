"""The shared discrepancy-workbook engine behind every comparison type.

Extracted (v0.10.0) from compare_highway_log.py — the approved TSMIS-vs-TSN
Highway Log workbook — and parameterized with a CompareSchema so the SAME
proven engine also builds the cross-environment comparisons (e.g. SSOR-prod
vs ARS-prod) for any report whose rows are "one key column + N data fields".
compare_highway_log delegates here with its original schema; its output is
regression-verified cell-for-cell identical to the pre-extraction engine, so
DO NOT change formula or label text here without re-running that check
(see CLAUDE.md — the per-route format is locked to the approved sample).

What the engine builds (sheet for sheet, same as the approved design):
  Summary       row counts, match status, per-field diff counts, live
                SELF-CHECK rows, notes
  Spot Check    one row audited field-by-field with an independent verdict
  Comparison    one row per (Route,) key + occurrence, document order;
                matched values shown, differences as "a ≠ b" in red
  Only in <A> / Only in <B>   every one-sided row, pulled onto its own tab
  Routes        (consolidated shape only) per-route coverage stats
  <A> / <B>     the two inputs + live "Key (helper)" lookup columns

Two flavors via mode= ("formulas" | "values" | "both"): formulas = every
cell a live Excel formula (consolidated ships calcMode=manual — ~2M formulas);
values = the same sheets/links/colors with the bulk precomputed by the same
Python mirror that powers the run summary, so the flavors can never disagree.

Inputs arrive as ROWS (the callers own file loading/validation): per-route
shape = [key, f1..fn]; consolidated shape = [route, key, f1..fn].

Streaming (write_only) workbook rules apply throughout: sheets created in
display order, freeze/widths/filter/CF set before rows, styled cells are
WriteOnlyCells. Console-free; progress via events.on_log, cancel honored at
the usual cadence, ConsolidateResult returned.
"""
import difflib
import math
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

import outcome
from comparison_contract import (
    CAPPED_FALLBACK_POLICY,
    CAPPED_PAIRING_ALGORITHM,
    SOURCE_PAIRING_ALGORITHM,
    CappedGroupDiagnostic,
    ComparisonCounts,
    PairingPair,
    PairingTrace,
    PhysicalIdentity,
    PhysicalKey,
    comparison_outcome_from_legacy,
    physical_identity_from_key,
)
from errors import RunCancelled

try:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

# Tint for a dittoed roadbed cell on the data sheets (a soft lavender, distinct
# from the diff/one-sided fills); only applied when a schema supplies a
# ditto_resolver (Highway Log), so other comparisons' data sheets are unchanged.
_DITTO_FILL = "E4DFEC"

from events import ConsolidateResult, Events

# Styling shared by all sheets (colors taken from the approved sample; the
# Only-in tab colors echo the Comparison sheet's yellow/blue row tints).
_DARK = "1F3864"            # header band / banners
_TAB_COLORS = {"summary": "808080", "spot": "7030A0", "comparison": "C00000",
               "routes": "ED7D31", "only_a": "BF8F00", "only_b": "2E75B6",
               "side_a": "4472C4", "side_b": "70AD47"}
_DIFF_MARK = " ≠ "          # presentation separator only; never owns diff truth

# E1 Med-Wid formula twin.  The optional suffix is deliberately an explicit
# printable-ASCII whitelist: Unicode digits/letters and control characters stay
# anomalous raw text.  Each Med-Wid source field gets these five hidden,
# versioned helper columns after Key(helper); formulas and values workbooks use
# identical physical geometry.
_MEDWID_SUFFIXES = "".join(chr(i) for i in range(0x21, 0x7F)
                            if chr(i) not in "0123456789.")
_MEDWID_HELPER_VERSION = "CMP_E1_MW_V1"
_MEDWID_HELPER_STAGES = ("TRIM", "CORE", "VALID", "MASK", "CANON")
_STATE_MASK_VERSION = "CMP_E1_STATE_V1"
_BUILD_SNAPSHOT_VERSION = "CMP_E2_SNAPSHOT_V1"
_BUILD_FRESH_VERSION = "CMP_E2_BUILD_FRESH_V1"
# Keep generated state formulas comfortably below Excel's hard 8,192-character
# ceiling.  The planner measures at Excel's maximum row number, so every formula
# actually emitted is no longer than the formula used to choose its chunk.
_STATE_FORMULA_TARGET = 7_600
_EXCEL_FORMULA_LIMIT = 8_192

_PROGRESS_EVERY = 2_500     # log + cancel-check cadence on big workbooks. Tight
                            # enough that the largest report (Intersection Detail,
                            # ~17k rows per sheet) never goes minutes without a
                            # progress line — long silences read as a hang and get
                            # cancelled. Also makes cancel land sooner (same check).

XL_MAX_ROWS, XL_MAX_COLS = 1_048_576, 16_384   # Excel worksheet hard limits


def excel_limit_error(n_rows, n_cols):
    """A user-safe message when a workbook would exceed Excel's worksheet
    limits, else None. Checked before writing so a too-large comparison fails
    cleanly instead of openpyxl raising mid-write (row cap) or silently
    dropping columns (column cap)."""
    if n_rows > XL_MAX_ROWS:
        return (f"This comparison needs {n_rows:,} rows, past Excel's "
                f"{XL_MAX_ROWS:,}-row limit. Compare a smaller scope (e.g. "
                "per-route files, or one report at a time).")
    if n_cols > XL_MAX_COLS:
        return (f"This report has {n_cols:,} columns, past Excel's "
                f"{XL_MAX_COLS:,}-column limit.")
    return None


@dataclass
class CompareSchema:
    """Everything report-specific the engine needs.

    The defaults reproduce the approved TSMIS-vs-TSN Highway Log wording —
    new comparison types override the data-shape fields and the side names.
    """
    report_name: str                  # "Highway Log" — titles/messages
    header: list                      # per-route columns: [key, f1..fn]
    side_a: str = "TSMIS"             # side A sheet/tab name (also in formulas)
    side_b: str = "TSN"               # side B sheet/tab name
    # The row-identity noun used in labels ("location" for the Highway Log,
    # "row" for generic reports, "route" for the Ramp Summary).
    id_noun: str = "location"
    id_noun_plural: str = "locations"
    pair_noun: str = ""               # the duplicate-key example noun in the
                                      # pairing note ("postmile"); "" = id_noun
    sides_noun: str = "systems"       # "both systems" / "both environments"
    medwid_fields: tuple = ()         # field NAMES normalized like Med Wid
    date_fields: tuple = ()           # field NAMES date-formatted on Spot Check
    data_widths: dict = field(default_factory=dict)   # field name -> data-sheet width
    cmp_widths: dict = field(default_factory=dict)    # field name -> comparison width
    scope_flat: str = "Per-route"     # Summary scope label, per-route shape
    scope_consolidated: str = "Consolidated (all routes)"
    one_sided_note_extra: str = ""    # appended to the yellow/blue note
    trim_note_extra: str = ""         # appended to the TRIM note
    # Which header column is the row-identity key (rows align/pair on it).
    # Default 0 = the first column, the original behavior. Reports whose first
    # column is coarse (e.g. cross-env Highway Sequence inherits County, which
    # repeats for hundreds of rows) point this at a granular column (postmile)
    # so rows align by identity instead of positionally within the coarse
    # group. The column stays in its display position everywhere — only the
    # identity used for alignment + the Comparison sheet's lead column change.
    key_field: int = 0
    # Optional per-report decoration (default None = the byte-identical original
    # output; only the Highway Log comparisons set these, so other comparisons
    # and the regression-locked fixtures are unchanged):
    #   header_comment — callable(label) -> openpyxl Comment | None, attached as a
    #     hover tooltip to each column-label header cell (Comparison / Only-in /
    #     data sheets).
    #   legend_writer  — callable(wb) run after the sheets, before save (e.g. to
    #     append a column-Legend sheet).
    header_comment: object = None
    legend_writer: object = None
    # When True, a cell whose value is a "+"-run ditto marker ('+', '++', '+++')
    # is NON-ASSERTING: it never counts as a difference against the other side
    # (its real value is compared on the paired roadbed's own row). Only the
    # Highway Log comparisons set this; for every other comparison it is False,
    # so the equality formula, the Spot Check verdict, and the Python mirrors are
    # byte-identical to the regression-locked output. See
    # docs/highway_log/comparison-study.md and highway_log_columns.is_ditto.
    ditto_nonasserting: bool = False
    # Optional DISPLAY-only resolver: callable(rows, has_route) ->
    # {row_index: {col_in_row: resolved_value}} for each dittoed cell. When set
    # (Highway Log), the data sheets keep the raw ditto in the cell (the
    # non-asserting diff depends on detecting the '+'-run) but tint it and attach
    # a comment with the paired-roadbed value, so a reviewer sees what each
    # `+`/`++` resolved to. None for every other comparison -> data sheets stay
    # byte-identical.
    ditto_resolver: object = None
    # Optional KEY normalizer: callable(row, off, key_field) -> scalar key, the
    # canonical identity token used for matching/alignment IN PLACE OF the raw
    # key-column value. It may return PhysicalKey only under uniform typed mode.
    # When None (every comparison except TSMIS-vs-TSN Highway Log), the raw
    # key-column string is used exactly as before, so the union, helper keys, and
    # MATCH lookups are byte-identical and the regression lock holds. The Highway
    # Log TSMIS-vs-TSN schemas set it to highway_log_columns.roadbed_canonical_
    # location so a divided segment's roadbed row keys identically whether the
    # source suffixes the Location (PDF/Excel) or dittos a block (TSN). The DATA
    # sheets still show each source's raw Location; only the key token is unified.
    key_normalizer: object = None
    # Optional CONTEXT fields: field NAMES that are SHOWN but never asserted on —
    # they never count as a difference and never get the ≠ mark / diff highlight.
    # Used for columns that exist on only one side (e.g. the TSN Ramp Detail DB
    # columns ADT / Ramp Type that the TSMIS report has no counterpart for): the
    # comparison still displays the value (coalescing the non-blank side) so a
    # reviewer sees it, but it contributes zero diff cells. Default () -> is_context
    # is always False, so every existing comparison (and the regression-locked
    # Route-1=969 Highway Log canary) is byte-identical.
    context_fields: tuple = ()
    # Optional EXTRA sheet writer: callable(wb, ctx) run after the standard sheets,
    # before save — like legend_writer but with a context dict ({rows_a, rows_b,
    # has_route, sc, side_a, side_b}) so a report can append a FAMILIAR-LAYOUT sheet
    # (e.g. the Ramp Summary category-count rollup) rendered from the compared rows.
    # Default None -> no extra sheet, byte-identical to before.
    extra_sheet_writer: object = None
    # Optional Summary-side invariant for a two-line familiar Report View.
    # Shape: (sheet_name, numeric Diffs column letter, physical rows per logical
    # record). The extra writer must create that sheet. Summary proves its
    # aggregate Diffs equals the generic Comparison total times the repeat.
    # This stays live in BOTH flavors and catches aggregate count drift. It is
    # not a value/per-field checksum: a same-count edit can leave it OK, so the
    # formulas Summary also labels Report View as a build-time snapshot.
    report_view_diff_check: tuple = ()
    # Source-file provenance (default "Source Files" companion sheet): per side, the
    # (export_prefix, consolidated_sheet, ext) used to name each row's per-route
    # source `<prefix>_route_<route>.<ext>` from that side's consolidated Route
    # column. `source_file_a` is the TSMIS/side-A input; `source_file_b` is set only
    # when side B is ALSO a per-route TSMIS export (the same-source PDF-vs-Excel /
    # cross-environment flavors). Empty () means that side has no per-route source
    # (e.g. the statewide TSN side). The run_files_compare / cross-env substrates
    # read these + compose the companion sheet — no per-report writer needed.
    source_file_a: tuple = ()
    source_file_b: tuple = ()

    @property
    def n_fields(self):
        return len(self.header) - 1

    @property
    def field_indices(self):
        """Header indices shown as data fields — every column except the key,
        in display order. key_field == 0 gives [1, 2, …, n] (the original
        behavior, so the default path is byte-identical)."""
        return [i for i in range(len(self.header)) if i != self.key_field]

    def is_medwid(self, field_idx):
        return self.header[field_idx] in self.medwid_fields

    def is_context(self, field_idx):
        """True for a non-asserting CONTEXT field (shown, never counted as a diff)."""
        return self.header[field_idx] in self.context_fields

    def sheet_names(self, has_route):
        names = ["Summary", "Spot Check", "Comparison"]
        if has_route:
            names.append("Routes")
        names += [f"Only in {self.side_a}", f"Only in {self.side_b}",
                  self.side_a, self.side_b,
                  "__CMP_E2_SNAPSHOT_A", "__CMP_E2_SNAPSHOT_B"]
        return names


def _sref(name):
    """`name` as an Excel formula sheet reference: bare when it's a plain
    identifier (keeps TSMIS/TSN formulas byte-identical to the approved
    sample), quoted otherwise ('SSOR-PROD', 'Only in TSMIS')."""
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name) \
            and not re.fullmatch(r"[A-Za-z]{1,3}\d{1,7}", name):
        return name
    return "'" + name.replace("'", "''") + "'"


# =============================================================================
# Keys + union (document-order diff merge)
# =============================================================================

def _key_text(v):
    """Identity text for a raw key cell (CMP-AUD-009).

    Canonicalize a numeric key the SAME way `_xl_trim` canonicalizes the
    compared value — an integral float is its integer (`5.0` -> `'5'`), a bool
    its literal — so a numeric-typed key aligns with its text twin (`5` / `"5"`)
    instead of splitting one physical row into two false one-sided rows. This is
    the alignment mirror of the value coercion `_xl_trim(5) == _xl_trim(5.0) ==
    "5"`, which rows never reached when their keys disagreed on type.

    Whitespace and case stay SIGNIFICANT: key IDENTITY must not be display-
    normalized (trim / collapse / casefold) the way values are — that would
    merge genuinely distinct keys. `' K '` and `'K'`, `''` and `' '` therefore
    remain distinct identities by design."""
    if type(v) is bool:
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def keys_for(rows, has_route, key_field=0, key_normalizer=None,
             is_cancelled=None):
    """[(route, key, occurrence), ...] in file order.

    Legacy per-route rows retain route ``""`` exactly. A PhysicalKey instead
    uses its validated canonical route on both workbook layouts so structured
    mode has one route authority. Occurrence repeats of the same (route, key)
    are numbered 1.., exactly like the sheets' live helper column.

    `key_field` picks WHICH header column is the identity key (default 0 = the
    first column, the original behavior). The raw row carries a leading Route
    when has_route, so the key sits at r[(1 if has_route else 0) + key_field].

    `key_normalizer` (opt-in; None = byte-identical original behavior) is a
    callable(row, off, key_field) returning the canonical identity token used
    IN PLACE OF the raw key-column string (the Highway Log roadbed key). It may
    return a PhysicalKey only when every key on both sides is migrated."""
    _raise_if_cancelled(is_cancelled)
    seen = {}
    out = []
    off = 1 if has_route else 0
    koff = off + key_field
    for row_no, r in enumerate(rows):
        if row_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        raw_route = r[0] if has_route else None
        route = "" if not has_route or raw_route is None else str(raw_route)
        physical_identity = physical_identity_from_key(r, off, key_field)
        if physical_identity is not None:
            # Keep the typed key object intact: its PhysicalIdentity, not its
            # source-visible str payload, owns equality and hashing.
            loc = r[koff]
        elif key_normalizer is not None:
            loc = key_normalizer(r, off, key_field)
        else:
            loc = "" if r[koff] is None else _key_text(r[koff])
        if physical_identity is None and isinstance(loc, PhysicalKey):
            physical_identity = loc.physical_identity
        if physical_identity is not None:
            canonical_route = physical_identity.canonical_components[0][1]
            if has_route and (type(raw_route) is not str
                              or raw_route != canonical_route):
                raise ValueError(
                    f"row {row_no + 1} outer route must exactly equal its "
                    f"PhysicalIdentity route {canonical_route!r}")
            # The canonical identity owns route grouping on every structured
            # path, including the route-less workbook layout.
            route = canonical_route
        k = (route, loc)
        seen[k] = seen.get(k, 0) + 1
        out.append((route, loc, seen[k]))
    _raise_if_cancelled(is_cancelled)
    return out


def _physical_identity(value):
    """The typed identity attached to a key, when this is the L0a path."""
    if not isinstance(value, PhysicalKey):
        return None
    identity = value.physical_identity
    return identity if isinstance(identity, PhysicalIdentity) else None


def _validate_uniform_physical_keys(keys_a, keys_b):
    """Reject partial migration and any second route authority before pairing."""
    sides = (("side A", keys_a), ("side B", keys_b))
    typed_present = any(
        _physical_identity(key[1]) is not None
        for _side, keys in sides for key in keys)
    if not typed_present:
        return
    for side, keys in sides:
        for index, key in enumerate(keys):
            identity = _physical_identity(key[1])
            if identity is None:
                raise ValueError(
                    "physical identity mode requires every key on both "
                    f"nonempty sides to be a PhysicalKey; {side} row "
                    f"{index + 1} is legacy")
            canonical_route = identity.canonical_components[0][1]
            if type(key[0]) is not str or key[0] != canonical_route:
                raise ValueError(
                    f"{side} row {index + 1} key route must exactly equal its "
                    f"PhysicalIdentity route {canonical_route!r}")


def _visible_key(value):
    """Workbook presentation for a key without flattening dictionary identity."""
    identity = _physical_identity(value)
    return identity.display if identity is not None else value


def published_key_text(sc, row, off=1):
    """The Comparison sheet's key cell for one LOADED row (CMP-AUD-208).

    An evidence adapter addresses a published row by asking the engine for the
    same presentation the engine wrote, rather than re-flattening key text of
    its own — so a caption, a highlight, and the published cell can never be
    about different rows.
    """
    return _visible_key(row[off + sc.key_field])


def _pairing_key_components(group, has_route):
    """Canonical trace components, preserving the legacy scalar trace shape."""
    identity = _physical_identity(group[1])
    if identity is not None:
        # Route is already part of the physical identity; do not duplicate the
        # outer grouping route in the auditable trace.
        return tuple(value for _name, value in identity.canonical_components)
    return ((str(group[0]), str(group[1])) if has_route
            else (str(group[1]),))


def union_keys(keys_t, keys_n, is_cancelled=None):
    """The union of the two key sequences in DOCUMENT order, grouped by route:
    side A's routes in side-A order (B-only routes appended in B order),
    and within each route a diff-style alignment of the two row sequences.

    Common keys appear exactly once (first position wins — a key can fall
    outside the aligner's 'equal' blocks when one file lists it out of
    sequence; seen in the field: TSMIS printed 059.739 after 059.759 while
    TSN kept it in order). The Excel MATCH lookups pair each union row with
    both files regardless of where it sits. Aligning per route keeps the
    matcher fast on consolidated inputs (50k+ rows)."""
    _raise_if_cancelled(is_cancelled)
    by_route_t, by_route_n = {}, {}
    routes = []
    for key_no, k in enumerate(keys_t):
        if key_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if k[0] not in by_route_t:
            by_route_t[k[0]] = []
            routes.append(k[0])
        by_route_t[k[0]].append(k)
    for key_no, k in enumerate(keys_n):
        if key_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if k[0] not in by_route_n:
            by_route_n[k[0]] = []
            if k[0] not in by_route_t:
                routes.append(k[0])
        by_route_n[k[0]].append(k)

    out = []
    seen = set()
    emitted_scans = 0

    def emit(keys, start=0, end=None):
        nonlocal emitted_scans
        stop = len(keys) if end is None else end
        for key_no in range(start, stop):
            if emitted_scans % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            emitted_scans += 1
            k = keys[key_no]
            if k not in seen:
                seen.add(k)
                out.append(k)

    for route in routes:
        _raise_if_cancelled(is_cancelled)
        seq_t = by_route_t.get(route, [])
        seq_n = by_route_n.get(route, [])
        if not seq_t or not seq_n:
            emit(seq_t or seq_n)
            _raise_if_cancelled(is_cancelled)
            continue
        # SequenceMatcher itself is an atomic stdlib call; bracket it and each
        # returned opcode so cancellation is observed at every safe boundary.
        sm = difflib.SequenceMatcher(None, seq_t, seq_n, autojunk=False)
        _raise_if_cancelled(is_cancelled)
        opcodes = sm.get_opcodes()
        _raise_if_cancelled(is_cancelled)
        for op, a0, a1, b0, b1 in opcodes:
            _raise_if_cancelled(is_cancelled)
            if op == "equal" or op == "delete":
                emit(seq_t, a0, a1)
            elif op == "insert":
                emit(seq_n, b0, b1)
            else:                       # replace: side-A block, then side-B block
                emit(seq_t, a0, a1)
                emit(seq_n, b0, b1)
            _raise_if_cancelled(is_cancelled)
        _raise_if_cancelled(is_cancelled)
    _raise_if_cancelled(is_cancelled)
    return out


def _opaque_helper_tokens(union, is_cancelled=None):
    """Injective workbook lookup tokens assigned in Comparison-row order.

    Identity components are intentionally absent from the token. Delimiter
    concatenation is not injective (``("R|X", "K")`` collides with
    ``("R", "X|K")``), while a versioned ordinal is both literal-safe and
    unique within the one workbook generation.
    """
    _raise_if_cancelled(is_cancelled)
    width = max(8, len(str(len(union))))
    tokens = {}
    for index, key in enumerate(union, start=1):
        if (index - 1) % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        tokens[key] = f"__{_HELPER_KEY_VERSION}_{index:0{width}d}"
    _raise_if_cancelled(is_cancelled)
    return tokens


# =============================================================================
# Python mirror of the workbook's comparison semantics (for the run summary;
# the formulas workbook recomputes everything live)
# =============================================================================

def normalize_value(v):
    """Canonicalize a freshly-LOADED cell so the comparison is type-stable and
    the two output flavors can't disagree. A real date/datetime is rendered to a
    fixed ISO string: Excel's TRIM of a live date is locale/number-format
    dependent and would diverge from Python's str(datetime) (e.g. "2/25/1976"
    vs "1976-02-25 00:00:00"), so the formulas flavor and the values flavor
    would compute different results for the SAME cell. Stringifying at load time
    means the engine only ever sees text — both flavors agree, and a date that
    is genuinely equal on both sides compares equal. Everything else (numbers,
    text) is returned unchanged. Callers (the loaders) apply this per cell."""
    # bool is a subclass of int.  Test its exact type first so only an actual
    # Boolean receives the TRUE/FALSE fold; numeric 1/0 and int subclasses stay
    # numeric.  Uppercase matches Excel's logical-cell display and gives the
    # Python/values flavor the same case-sensitive text the formulas flavor sees.
    if type(v) is bool:
        return "TRUE" if v else "FALSE"
    if isinstance(v, datetime):
        if (v.hour, v.minute, v.second, v.microsecond) == (0, 0, 0, 0):
            return v.date().isoformat()
        return v.isoformat(sep=" ")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):       # a time-only cell would otherwise reach the
        return v.isoformat()      # data sheet typed (Excel) vs str() (values)
    return v


def _xl_trim(v):
    """Excel TRIM: text form, edge U+0020 stripped, U+0020 runs collapsed.

    Tabs, CR/LF, NBSP, and every other Unicode whitespace character are data,
    not globally normalized whitespace.  The explicit Boolean branch also
    makes this public policy seam safe for callers that have not yet applied
    ``normalize_value`` themselves.
    """
    if v is None:
        return ""
    if type(v) is bool:
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return re.sub(" +", " ", str(v)).strip(" ")


def _is_plus_run(v):
    """True if `v` is a Highway Log ditto marker — a non-empty run of only '+'
    ('+', '++', '+++'). Mirrors highway_log_columns.is_ditto, kept local so the
    generic engine carries no Highway-Log import; gated by
    CompareSchema.ditto_nonasserting so it is inert for every other comparison."""
    if v is None:
        return False
    # Excel TRIM semantics remove/collapse ordinary ASCII spaces only.  Tabs,
    # CR/LF, NBSP, and other Unicode whitespace remain comparison data and
    # must not turn a non-marker value into a Highway Log ditto marker.
    s = _xl_trim(v)
    return bool(s) and set(s) == {"+"}


def _medwid_norm(t):
    """Canonical narrow unsigned Med-Wid text without numeric coercion.

    Accepted grammar is ASCII ``digits[.digits][suffix]``, where suffix is one
    printable ASCII character other than a digit or dot.  Leading integer
    zeros and trailing fractional zeros are insignificant; the optional suffix
    is retained byte-for-byte, so suffix case remains significant.  The
    printable-ASCII boundary is deliberately fail-closed: Unicode digits,
    non-ASCII lookalikes, and control characters remain anomalous raw text.
    Signs, leading decimals, exponents, Unicode digits, and every other shape
    stay raw.  String canonicalization is decimal-exact for arbitrarily long
    inputs and never passes through binary ``float`` or Excel ``VALUE``.
    """
    match = re.fullmatch(r"([0-9]+)(?:\.([0-9]+))?(.)?", t)
    if match is None:
        return t
    whole, fraction, suffix = match.groups()
    if suffix is not None and suffix not in _MEDWID_SUFFIXES:
        return t
    whole = whole.lstrip("0") or "0"
    if fraction is not None:
        fraction = fraction.rstrip("0")
    number = whole + (f".{fraction}" if fraction else "")
    return number + (suffix or "")


@dataclass(frozen=True)
class ComparedCell:
    """One canonical compared-cell decision shared by every Python consumer.

    ``raw_a``/``raw_b`` are the values supplied to the engine.  ``display_a``
    and ``display_b`` are their ASCII-TRIMmed forms; ``normalized_a`` and
    ``normalized_b`` are the exact equality operands (including Med-Wid when
    configured).  ``equal`` always describes those operands, while
    ``asserting`` says whether that equality contributes to discrepancy truth.
    ``display`` is the final values-workbook text.

    Existing evidence adapters unpack ``(va, vb, verdict)``.  Iteration keeps
    that compatibility, returning ``None`` as the third item for a
    non-asserting cell even though this typed result still records ``equal``.
    """

    raw_a: object
    raw_b: object
    normalized_a: str
    normalized_b: str
    asserting: bool
    equal: bool
    display_a: str
    display_b: str
    display: str

    @property
    def verdict(self):
        return self.equal if self.asserting else None

    @property
    def state_code(self):
        """Compact compared-cell truth used by every workbook consumer.

        ``E`` and ``D`` are asserting equal/different decisions; ``N`` is a
        displayed but non-asserting context/ditto cell.  One-sided ``U`` is a
        row-level decision and is assigned by :func:`count_diffs`.
        """
        if not self.asserting:
            return "N"
        return "E" if self.equal else "D"

    def __iter__(self):
        yield self.display_a
        yield self.display_b
        yield self.verdict


# =============================================================================
# Similarity pairing for duplicate keys
# =============================================================================
# Rows that share a key (same Route + key field) used to pair by FILE ORDER:
# first-with-first, second-with-second. When a key legitimately repeats (two
# segments at the same postmile), that mis-paired a row with the wrong twin and
# reported phantom differences — the row that actually matched the OTHER side's
# second instance looked like a difference. Instead, within each duplicate group
# present on BOTH sides, pair the rows that are actually the MOST ALIKE (fewest
# differing fields) and give matched pairs the same occurrence #. The optimal
# assignment's total difference count is <= any positional assignment's (file
# order is just one assignment), so this can only REMOVE phantom diffs, never add
# one.

_PAIR_GROUP_CAP = 100_000     # exact rectangular assignment through this product
_HELPER_KEY_VERSION = "CMP_E2_KEY_V1"


def _raise_if_cancelled(is_cancelled):
    """Abort an in-flight pairing phase without returning partial identity.

    The callback is deliberately a bare predicate instead of an ``Events``
    object so the exact assignment helper remains independently testable.  A
    false predicate never changes solver state or tie ordering; a true one is
    represented by the same ``RunCancelled`` signal used by the other engines.
    """
    if is_cancelled is not None and is_cancelled():
        raise RunCancelled()


def compared_cell(sc, f, rt, rn, off):
    """Return the one typed equality/display decision for field ``f``.

    Ordinary equality is exact and case-sensitive.  Context and configured
    ditto cells retain their normalized equality for diagnostics but are marked
    non-asserting, so no Python truth consumer needs a sentinel or display-text
    scan.  ``ComparedCell.__iter__`` preserves legacy tuple unpacking.
    """
    raw_a, raw_b = rt[f + off], rn[f + off]
    display_a, display_b = _xl_trim(raw_a), _xl_trim(raw_b)
    if sc.is_medwid(f):
        normalized_a = _medwid_norm(display_a)
        normalized_b = _medwid_norm(display_b)
    else:
        normalized_a, normalized_b = display_a, display_b
    equal = normalized_a == normalized_b
    is_context = sc.is_context(f)
    is_ditto = (sc.ditto_nonasserting
                and (_is_plus_run(display_a) or _is_plus_run(display_b)))
    asserting = not (is_context or is_ditto)
    if is_context:
        display = display_a if display_a else display_b
    elif is_ditto or equal:
        display = display_a
    else:
        display = (f"{display_a or '(blank)'}{_DIFF_MARK}"
                   f"{display_b or '(blank)'}")
    return ComparedCell(
        raw_a=raw_a,
        raw_b=raw_b,
        normalized_a=normalized_a,
        normalized_b=normalized_b,
        asserting=asserting,
        equal=equal,
        display_a=display_a,
        display_b=display_b,
        display=display,
    )


def _row_diff_count(sc, rt, rn, off):
    """Differing ASSERTING compared-fields between two rows, using the
    comparison's own normalization (TRIM + Med Wid) — identical to the per-row
    diff the workbook counts. This is the verdict vocabulary (the persisted
    per-pair cost) and the capped path's diagonal diagnostic; the within-cap
    assignment objective is ``_pair_cost_components``."""
    d = 0
    for f in sc.field_indices:
        cell = compared_cell(sc, f, rt, rn, off)
        if cell.asserting and not cell.equal:
            d += 1
    return d


def _char_distance(a, b, cache):
    """Exact Levenshtein distance between two normalized cell texts, memoized
    per duplicate-group build (the same value pair recurs across a group's
    matrix; the distance is symmetric, so the key is order-normalized).
    Mirrors the Stage-8 oracle's reference implementation."""
    if a == b:
        return 0
    key = (a, b) if a <= b else (b, a)
    hit = cache.get(key)
    if hit is not None:
        return hit
    left, right = key
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for left_index, left_character in enumerate(left, 1):
        current = [left_index]
        for right_index, right_character in enumerate(right, 1):
            current.append(min(
                current[-1] + 1,
                previous[right_index] + 1,
                previous[right_index - 1]
                + (left_character != right_character),
            ))
        previous = current
    cache[key] = previous[-1]
    return previous[-1]


def _pair_cost_components(sc, rt, rn, off, char_cache):
    """One candidate pair's cost components in a single compared-cell pass:
    ``(asserting diffs, all-compared-field diffs, summed char edit distance)``.

    The asserting count stays the persisted ``PairingPair.cost`` — the verdict
    vocabulary. The other two feed the source-identity assignment objective
    (CMP-AUD-220, approved 2026-07-16): context and ditto cells distinguish
    which physical occurrences correspond without ever asserting a
    difference, exactly the Stage-8 oracle's all-field reading."""
    asserting = compared = chars = 0
    for f in sc.field_indices:
        cell = compared_cell(sc, f, rt, rn, off)
        if cell.equal:
            continue
        compared += 1
        chars += _char_distance(
            cell.normalized_a, cell.normalized_b, char_cache)
        if cell.asserting:
            asserting += 1
    return asserting, compared, chars


def _min_cost_pairs(cost, is_cancelled=None):
    """Return the exact rectangular minimum-cost assignment.

    Every member of the smaller side is matched to a distinct member of the
    larger side. When several assignments have the same minimum scalar cost,
    the lexicographically smallest smaller-side assignment vector wins; side A
    is the smaller side when dimensions are equal. The Hungarian solver is
    genuinely rectangular (``O(min(n,m)^2 * max(n,m))``), so a 1-by-100,000
    boundary group never expands to a square matrix.

    The matrix contract is deliberately strict. Silent acceptance of an empty,
    ragged, Boolean, fractional, or negative cost matrix would make a pairing
    trace look authoritative when the objective was malformed.
    """
    _raise_if_cancelled(is_cancelled)
    if not isinstance(cost, (list, tuple)) or not cost:
        raise ValueError("assignment matrix must contain at least one row")
    if not isinstance(cost[0], (list, tuple)) or not cost[0]:
        raise ValueError("assignment matrix must contain at least one column")
    nr, nc = len(cost), len(cost[0])
    checked_cells = 0
    for row_no, row in enumerate(cost):
        if row_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if not isinstance(row, (list, tuple)) or len(row) != nc:
            raise ValueError("assignment matrix is ragged")
        for value in row:
            if checked_cells % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            checked_cells += 1
            if type(value) is not int or value < 0:
                raise ValueError(
                    "assignment costs must be non-negative integers")

    flip = nr > nc
    _raise_if_cancelled(is_cancelled)
    if not flip:
        oriented = cost
    else:
        transposed = []
        moved_cells = 0
        for b in range(nc):
            transposed_row = []
            for a in range(nr):
                if moved_cells % _PROGRESS_EVERY == 0:
                    _raise_if_cancelled(is_cancelled)
                moved_cells += 1
                transposed_row.append(cost[a][b])
            transposed.append(tuple(transposed_row))
        oriented = tuple(transposed)
    n, m = len(oriented), len(oriented[0])

    # Encode the two-level objective as one exact integer objective. The
    # primary multiplier dominates every possible base-(m+1) assignment-vector
    # suffix, so scalar cost is minimized first and the vector second.
    base = m + 1
    primary = base ** n
    lex_weights = tuple(base ** (n - 1 - i) for i in range(n))
    _raise_if_cancelled(is_cancelled)

    def encoded(i, j):
        return oriented[i][j] * primary + j * lex_weights[i]

    # Rectangular Hungarian potentials, with rows guaranteed <= columns.
    u = [0] * (n + 1)
    v = [0] * (m + 1)
    p = [0] * (m + 1)
    way = [0] * (m + 1)
    for i in range(1, n + 1):
        _raise_if_cancelled(is_cancelled)
        p[0] = i
        j0 = 0
        minv = [None] * (m + 1)
        used = [False] * (m + 1)
        while True:
            _raise_if_cancelled(is_cancelled)
            used[j0] = True
            i0 = p[j0]
            delta = None
            j1 = 0
            for j in range(1, m + 1):
                if j % _PROGRESS_EVERY == 0:
                    _raise_if_cancelled(is_cancelled)
                if used[j]:
                    continue
                cur = encoded(i0 - 1, j - 1) - u[i0] - v[j]
                if minv[j] is None or cur < minv[j]:
                    minv[j] = cur
                    way[j] = j0
                if (delta is None or minv[j] < delta
                        or (minv[j] == delta and j < j1)):
                    delta = minv[j]
                    j1 = j
            if delta is None:
                raise RuntimeError(
                    "rectangular assignment has no augmenting column")
            for j in range(m + 1):
                if j % _PROGRESS_EVERY == 0:
                    _raise_if_cancelled(is_cancelled)
                if used[j]:
                    u[p[j]] += delta
                    v[j] -= delta
                elif minv[j] is not None:
                    minv[j] -= delta
            j0 = j1
            if p[j0] == 0:
                break
        while True:
            j1 = way[j0]
            p[j0] = p[j1]
            j0 = j1
            if j0 == 0:
                break

    _raise_if_cancelled(is_cancelled)
    assignment = [-1] * n
    for j in range(1, m + 1):
        if p[j]:
            assignment[p[j] - 1] = j - 1
    if (any(value < 0 for value in assignment)
            or len(set(assignment)) != n):
        raise RuntimeError("rectangular assignment is incomplete")

    pairs = ([(i, j) for i, j in enumerate(assignment)] if not flip
             else [(j, i) for i, j in enumerate(assignment)])
    total = sum(cost[a][b] for a, b in pairs)
    positional = sum(cost[i][i] for i in range(min(nr, nc)))
    if total > positional:
        raise AssertionError(
            "exact assignment is worse than positional pairing")
    _raise_if_cancelled(is_cancelled)
    return pairs


@dataclass(frozen=True)
class OccurrencePairing:
    """Occurrence-renumbering result plus independently auditable evidence.

    Iteration preserves the historical ``keys_a, keys_b = result`` call seam;
    typed consumers use the named fields and cannot lose pairing quality.
    """

    keys_a: tuple
    keys_b: tuple
    pairing_trace: tuple
    pairing_quality: str
    capped_group_diagnostics: tuple

    @property
    def keys_t(self):
        return self.keys_a

    @property
    def keys_n(self):
        return self.keys_b

    def __iter__(self):
        if self.pairing_quality != "exact":
            raise ValueError(
                "capped occurrence pairing must be consumed through typed "
                "keys/quality/diagnostic fields")
        yield self.keys_a
        yield self.keys_b


def pair_occurrences_by_similarity(sc, rows_t, rows_n, keys_t, keys_n,
                                   has_route, events=None):
    """Re-number the occurrence component of DUPLICATE keys so that, within each
    (route, key) group present on BOTH sides, rows pair by SOURCE IDENTITY
    instead of by file order: the assignment minimizes the lexicographic
    (all-compared-field diff count, summed character edit distance, |position
    gap|) tuple, so context and ditto cells help decide WHICH physical
    occurrences correspond while verdicts/counts stay asserted-only
    (CMP-AUD-220, the approved D3 assignment/verdict split). Matched
    pairs are numbered 1.. in side-A file order; the larger side's leftovers get
    higher, side-unique occurrence numbers (so they stay one-sided). Groups with
    no duplicate, or a key on only one side, are untouched (occurrence # = file
    order, exactly as before).

    Within the product cap, every duplicate group is assigned exactly. Above
    the cap, positional pairs remain useful diagnostic output, but typed quality
    is ``capped`` and the caller must make the comparison partial and
    non-certifying. No above-cap full matrix is allocated.
    """
    off = 1 if has_route else 0
    is_cancelled = events.is_cancelled if events is not None else None
    _raise_if_cancelled(is_cancelled)
    _validate_uniform_physical_keys(keys_t, keys_n)
    grp_t, grp_n = {}, {}
    for i, k in enumerate(keys_t):
        if i % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        grp_t.setdefault((k[0], k[1]), []).append(i)
    for i, k in enumerate(keys_n):
        if i % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        grp_n.setdefault((k[0], k[1]), []).append(i)

    out_t, out_n = list(keys_t), list(keys_n)
    traces = []
    capped_diagnostics = []
    for grp, tis in grp_t.items():
        _raise_if_cancelled(is_cancelled)
        nis = grp_n.get(grp)
        if not nis or (len(tis) == 1 and len(nis) == 1):
            continue                        # key on one side only, or no duplicate
        na, nb = len(tis), len(nis)
        matrix_cells = na * nb
        smaller_side = "a" if na <= nb else "b"
        key_components = _pairing_key_components(grp, has_route)
        positional_pairs = [(i, i) for i in range(min(na, nb))]
        if matrix_cells > _PAIR_GROUP_CAP:
            # Observe only the diagonal used by the positional fallback; never
            # allocate or scan the prohibited full matrix.
            pairs = positional_pairs
            pair_costs = []
            for a, b in pairs:
                _raise_if_cancelled(is_cancelled)
                pair_costs.append(
                    _row_diff_count(
                        sc, rows_t[tis[a]], rows_n[nis[b]], off))
            assignment = tuple(range(min(na, nb)))
            total = sum(pair_costs)
            positional_cost = total
            pair_objectives = {}
            objective_total = None
            objective_positional = None
            quality = "capped"
            exact = False
            algorithm = CAPPED_PAIRING_ALGORITHM
            capped_diagnostics.append(CappedGroupDiagnostic(
                key_components=key_components,
                side_a_size=na,
                side_b_size=nb,
                matrix_cells=matrix_cells,
                cap=_PAIR_GROUP_CAP,
                fallback_policy=CAPPED_FALLBACK_POLICY,
                fallback_cost=total,
            ))
        else:
            asserting = []
            diffs = []
            chars = []
            char_cache = {}
            for ti in tis:
                asserting_row, diffs_row, chars_row = [], [], []
                for ni in nis:
                    _raise_if_cancelled(is_cancelled)
                    one = _pair_cost_components(
                        sc, rows_t[ti], rows_n[ni], off, char_cache)
                    asserting_row.append(one[0])
                    diffs_row.append(one[1])
                    chars_row.append(one[2])
                asserting.append(asserting_row)
                diffs.append(diffs_row)
                chars.append(chars_row)
            # Encode the lexicographic source-identity objective — (all-
            # compared-field diffs, char edit distance, |position gap|) — as
            # one exact integer per cell (CMP-AUD-220). Each weight strictly
            # bounds every feasible assignment's lower components, so scalar
            # order equals tuple order; _min_cost_pairs then appends the
            # approved smallest-smaller-side-vector tie rule (D3.2) unchanged.
            smaller_size = min(na, nb)
            position_weight = smaller_size * (max(na, nb) - 1) + 1
            max_chars = max(max(row) for row in chars)
            chars_weight = (smaller_size * max_chars + 1) * position_weight
            _raise_if_cancelled(is_cancelled)
            cost = [
                [diffs[a][b] * chars_weight + chars[a][b] * position_weight
                 + abs(a - b) for b in range(nb)]
                for a in range(na)
            ]
            pairs = _min_cost_pairs(
                cost, is_cancelled)         # local (side-A, side-B) indices
            pair_costs = [asserting[a][b] for a, b in pairs]
            pair_objectives = {
                (a, b): (diffs[a][b], chars[a][b], abs(a - b))
                for a, b in pairs}
            if smaller_side == "a":
                assignment = tuple(
                    b for a, b in sorted(pairs, key=lambda pair: pair[0]))
            else:
                assignment = tuple(
                    a for a, b in sorted(pairs, key=lambda pair: pair[1]))
            total = sum(pair_costs)
            positional_cost = sum(
                asserting[i][i] for i in range(smaller_size))
            objective_total = tuple(
                sum(pair_objectives[pair][index] for pair in pairs)
                for index in range(3))
            objective_positional = (
                sum(diffs[i][i] for i in range(smaller_size)),
                sum(chars[i][i] for i in range(smaller_size)),
                0,
            )
            # The minimized quantity is the objective; the asserted total may
            # legitimately exceed file order when context identity demands it.
            if objective_total > objective_positional:
                raise AssertionError(
                    "exact duplicate pairing is worse than file order")
            quality = "exact"
            exact = True
            algorithm = SOURCE_PAIRING_ALGORITHM

        cost_by_pair = {}
        for pair_no, pair in enumerate(pairs):
            if pair_no % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            cost_by_pair[pair] = pair_costs[pair_no]
        ordered_pairs = (sorted(pairs, key=lambda pair: pair[0])
                         if smaller_side == "a" else
                         sorted(pairs, key=lambda pair: pair[1]))
        _raise_if_cancelled(is_cancelled)
        typed_pairs = []
        for pair_no, (a, b) in enumerate(ordered_pairs):
            if pair_no % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            typed_pairs.append(PairingPair(
                side_a_index=tis[a],
                side_b_index=nis[b],
                cost=cost_by_pair[(a, b)],
                objective=pair_objectives.get((a, b)),
            ))
        typed_pairs = tuple(typed_pairs)
        _raise_if_cancelled(is_cancelled)
        traces.append(PairingTrace(
            key_components=key_components,
            side_a_size=na,
            side_b_size=nb,
            matrix_cells=matrix_cells,
            side_a_indices=tuple(tis),
            side_b_indices=tuple(nis),
            smaller_side=smaller_side,
            assignment_vector=assignment,
            pairs=typed_pairs,
            total_cost=total,
            positional_cost=positional_cost,
            algorithm=algorithm,
            exact=exact,
            quality=quality,
            objective_total=objective_total,
            objective_positional=objective_positional,
        ))
        matched_t = {a for a, _ in pairs}
        matched_n = {b for _, b in pairs}
        occ = 0
        for pair_no, (a, b) in enumerate(
                sorted(pairs, key=lambda ab: tis[ab[0]])):
            if pair_no % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            occ += 1
            # Preserve each side's own key object. Equal PhysicalKey objects
            # may carry different lossless raw claims; copying the group key
            # from side A onto side B destroyed that evidence.
            out_t[tis[a]] = (*keys_t[tis[a]][:2], occ)
            out_n[nis[b]] = (*keys_n[nis[b]][:2], occ)
        extra = occ
        for a in range(len(tis)):           # unmatched side-A rows, file order
            if a % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            if a not in matched_t:
                extra += 1
                out_t[tis[a]] = (*keys_t[tis[a]][:2], extra)
        extra = occ
        for b in range(len(nis)):           # unmatched side-B rows, file order
            if b % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            if b not in matched_n:
                extra += 1
                out_n[nis[b]] = (*keys_n[nis[b]][:2], extra)
    _raise_if_cancelled(is_cancelled)
    if capped_diagnostics and events is not None:
        events.on_log(
            f"  Warning: {len(capped_diagnostics)} duplicate key group(s) "
            "exceeded the exact-pairing cap. Positional counts were retained "
            "for diagnosis, but this comparison is partial and cannot certify "
            "a match.")
    _raise_if_cancelled(is_cancelled)
    return OccurrencePairing(
        keys_a=tuple(out_t),
        keys_b=tuple(out_n),
        pairing_trace=tuple(traces),
        pairing_quality="capped" if capped_diagnostics else "exact",
        capped_group_diagnostics=tuple(capped_diagnostics),
    )


def count_diffs(sc, rows_t, rows_n, keys_t, keys_n, union, has_route,
                is_cancelled=None):
    """Counts matching what the workbook's formulas will compute: overall
    totals, per-field difference counts, per-route aggregates (consolidated),
    and the FIRST matched-with-differences Comparison row (the Spot Check
    sheet's default). The same numbers back the run summary AND become the
    literal cells of the values workbook, so the two output flavors can
    never disagree."""
    _raise_if_cancelled(is_cancelled)
    off = 1 if has_route else 0          # data fields start after Route
    by_t = {}
    for i, k in enumerate(keys_t):
        if i % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        by_t[k] = rows_t[i]
    by_n = {}
    for i, k in enumerate(keys_n):
        if i % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        by_n[k] = rows_n[i]
    both = t_only = n_only = diff_rows = identical = diff_cells = 0
    asserted_cells = context_cells = 0
    first_diff_row = None
    field_diffs = {f: 0 for f in sc.field_indices}
    state_masks = {}
    route = {}                           # consolidated: per-route aggregates

    def rstat(rid):
        return route.setdefault(rid, {"t_rows": 0, "n_rows": 0, "locs": 0,
                                      "matched": 0, "withdiffs": 0, "cells": 0})
    if has_route:
        for i, k in enumerate(keys_t):
            if i % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            rstat(k[0])["t_rows"] += 1
        for i, k in enumerate(keys_n):
            if i % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            rstat(k[0])["n_rows"] += 1
    compared_fields = 0
    for i, k in enumerate(union):
        if i % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        rt, rn = by_t.get(k), by_n.get(k)
        rs = rstat(k[0]) if has_route else None
        if rs is not None:
            rs["locs"] += 1
        if rt is None:
            n_only += 1
            state_masks[k] = "U" * len(sc.field_indices)
            continue
        if rn is None:
            t_only += 1
            state_masks[k] = "U" * len(sc.field_indices)
            continue
        both += 1
        if rs is not None:
            rs["matched"] += 1
        row_diffs = 0
        state_codes = []
        for f in sc.field_indices:                # every column but the key
            if compared_fields % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            compared_fields += 1
            cell = compared_cell(sc, f, rt, rn, off)
            state_codes.append(cell.state_code)
            if cell.asserting:
                asserted_cells += 1
            else:
                context_cells += 1
            if cell.asserting and not cell.equal:
                row_diffs += 1
                field_diffs[f] += 1
        state_masks[k] = "".join(state_codes)
        diff_cells += row_diffs
        if rs is not None:
            rs["cells"] += row_diffs
            if row_diffs:
                rs["withdiffs"] += 1
        if row_diffs:
            diff_rows += 1
            if first_diff_row is None:
                first_diff_row = i + 2
        else:
            identical += 1
    _raise_if_cancelled(is_cancelled)
    return {"both": both, "t_only": t_only, "n_only": n_only,
            "diff_rows": diff_rows, "identical": identical,
            "diff_cells": diff_cells, "first_diff_row": first_diff_row,
            "field_diffs": field_diffs, "route": route,
            "asserted_cells": asserted_cells, "context_cells": context_cells,
            "state_masks": state_masks}


# =============================================================================
# Layout: column geometry for the two input shapes
# =============================================================================

def _snapshot_cell_text(ref):
    """Exact raw-cell text for build-snapshot freshness comparisons."""
    return f'IF(ISBLANK({ref}),"",{ref}&"")'


def _build_fresh_formula(lay, side, row, columns):
    """Live proof that source/helper cells still equal their build snapshot."""
    snapshot = _sref(lay.snapshot_sheet(side))
    checks = []
    for column in columns:
        current = f"${column}{row}"
        expected = f"{snapshot}!${column}{row}"
        checks.append(
            f"EXACT({_snapshot_cell_text(current)},"
            f"{_snapshot_cell_text(expected)})")
    return f'=IF(AND({",".join(checks)}),"OK","STALE")'


def _build_tail_formula(row, columns):
    """Detect data/helper content appended below the build-time row universe."""
    ranges = [f"${column}${row}:${column}${XL_MAX_ROWS}" for column in columns]
    return f'=IF(COUNTA({",".join(ranges)})=0,"END","STALE")'

class _Layout:
    """Column letters for both workbook shapes.

    per-route:    data sheets  Comparison row=A (back-link), key=B,
                  fields C.., key helper, then hidden Med-Wid helpers
    consolidated: data sheets  Comparison row=A (back-link), Route=B,
                  key=C, fields D.., key helper, then hidden Med-Wid helpers
    """

    def __init__(self, sc, has_route):
        self.sc = sc
        self.has_route = has_route
        self.off = 1 if has_route else 0
        self.n_fields = sc.n_fields
        # The identity key column (header index) + the fields shown on the
        # Comparison sheet (every other column, in display order). key_field
        # == 0 reproduces the original [1..n] field order exactly.
        self.key_field = sc.key_field
        self.field_indices = sc.field_indices
        self._field_pos = {f: pos for pos, f in enumerate(self.field_indices)}
        # statuses as they appear in cells/formulas/CF
        self.only_a = f"{sc.side_a} only"
        self.only_b = f"{sc.side_b} only"
        # data sheets: a leading "Comparison row" back-link column, then the
        # input's columns (ALL of them, in their original order — the key
        # column stays in place there), then the key helper. Hidden Med-Wid
        # normalization helpers, when needed, are appended after it below.
        self.data_header = (["Route"] if has_route else []) + list(sc.header)
        self.back_col = "A"                                  # back-link column
        self.route_data_col = "B" if has_route else None     # Route on data sheets
        self.key_col = get_column_letter(len(self.data_header) + 2)
        self.data_last_col = get_column_letter(len(self.data_header) + 1)
        # Hidden Med-Wid helpers are APPENDED after the existing Key(helper), so
        # every visible/source/back-link/key coordinate above stays unchanged.
        # The stable schema index, not a possibly duplicated display label,
        # makes each internal header unique.
        self.medwid_field_indices = tuple(
            f for f in self.field_indices if sc.is_medwid(f))
        self._medwid_helper_cols = {}
        self.medwid_helper_headers = []
        helper_col_idx = len(self.data_header) + 3
        for f in self.medwid_field_indices:
            token = f"F{f:03d}"
            stage_cols = {}
            for stage in _MEDWID_HELPER_STAGES:
                stage_cols[stage] = get_column_letter(helper_col_idx)
                self.medwid_helper_headers.append(
                    f"__{_MEDWID_HELPER_VERSION}_{token}_{stage}")
                helper_col_idx += 1
            self._medwid_helper_cols[f] = stage_cols
        # E2 correctness lock: every source/helper cell has an immutable copy on
        # a very-hidden snapshot sheet. Live chunk formulas on each data sheet
        # compare the current cells to that snapshot. Any post-build source edit
        # therefore makes Summary non-certifying until the workbook is rebuilt;
        # static helper identities and duplicate assignments can never stay green.
        self.snapshot_sheets = {
            sc.side_a: "__CMP_E2_SNAPSHOT_A",
            sc.side_b: "__CMP_E2_SNAPSHOT_B",
        }
        self.snapshot_physical_n_cols = helper_col_idx - 1
        self.snapshot_physical_last_col = get_column_letter(
            self.snapshot_physical_n_cols)
        snapshot_columns = [get_column_letter(index) for index in range(
            2, self.snapshot_physical_n_cols + 1)]
        self.build_fresh_chunks = []
        current = []
        for column in snapshot_columns:
            single = _build_fresh_formula(
                self, sc.side_a, XL_MAX_ROWS, [column])
            if len(single) > _EXCEL_FORMULA_LIMIT:
                raise ValueError(
                    f"Build-snapshot formula for data column {column} exceeds "
                    f"Excel's {_EXCEL_FORMULA_LIMIT:,}-character limit.")
            candidate = current + [column]
            formula = _build_fresh_formula(
                self, sc.side_a, XL_MAX_ROWS, candidate)
            if current and len(formula) > _STATE_FORMULA_TARGET:
                self.build_fresh_chunks.append(tuple(current))
                current = [column]
            else:
                current = candidate
        if current:
            self.build_fresh_chunks.append(tuple(current))
        self.build_fresh_headers = []
        self.build_fresh_cols = []
        for chunk_no, columns in enumerate(self.build_fresh_chunks, start=1):
            col = get_column_letter(helper_col_idx)
            self.build_fresh_cols.append(col)
            self.build_fresh_headers.append(
                f"__{_BUILD_FRESH_VERSION}_C{chunk_no:03d}_"
                f"{columns[0]}_{columns[-1]}")
            helper_col_idx += 1
        self.physical_n_cols = helper_col_idx - 1
        self.physical_last_col = get_column_letter(self.physical_n_cols)
        # K:M are the independent Spot-Check state/display twins.  Med-Wid's
        # staged T:AC block remains farther right when present.
        self.spot_physical_n_cols = 29 if self.medwid_field_indices else 13
        # comparison sheet: the key column leads (pulled to the identity slot),
        # the remaining columns follow as fields in display order
        self.id_headers = ((["Route"] if has_route else [])
                           + [sc.header[self.key_field], "#", f"{sc.side_a} Row",
                              f"{sc.side_b} Row", "Status", "Diffs"])
        self.f0 = len(self.id_headers) + 1            # first field column index
        self.first_field_col = get_column_letter(self.f0)
        self.last_field_col = get_column_letter(self.f0 + self.n_fields - 1)
        c = [get_column_letter(i + 1) for i in range(len(self.id_headers))]
        if has_route:
            (self.c_route, self.c_loc, self.c_occ, self.c_trow, self.c_nrow,
             self.c_status, self.c_diffs) = c
        else:
            self.c_route = None
            (self.c_loc, self.c_occ, self.c_trow, self.c_nrow,
             self.c_status, self.c_diffs) = c

        # Hidden Comparison state masks follow every visible field.  Chunking is
        # planned from the live formula length at Excel's maximum row number;
        # fields stay in display order, so each mask character has one stable
        # positional meaning in both formula and values workbooks.
        self.state_chunks = []
        self._state_field_location = {}
        state_col_idx = self.f0 + self.n_fields
        state_pos = 0
        planned_state_chunks = _plan_state_chunk_fields(self)
        planned_last_col = (state_col_idx + len(planned_state_chunks) - 1
                            if planned_state_chunks
                            else self.f0 + self.n_fields - 1)
        state_limit_err = excel_limit_error(1, planned_last_col)
        if state_limit_err:
            raise ValueError(state_limit_err)
        for chunk_no, fields in enumerate(planned_state_chunks, start=1):
            col_idx = state_col_idx + chunk_no - 1
            col = get_column_letter(col_idx)
            start = state_pos
            end = start + len(fields) - 1
            header = (f"__{_STATE_MASK_VERSION}_C{chunk_no:03d}_"
                      f"P{start:04d}_P{end:04d}")
            chunk = {"fields": tuple(fields), "col_idx": col_idx, "col": col,
                     "header": header, "start": start, "end": end}
            self.state_chunks.append(chunk)
            for offset, field_idx in enumerate(fields, start=1):
                self._state_field_location[field_idx] = (chunk, offset)
            state_pos = end + 1
        last_state_col_idx = (
            self.state_chunks[-1]["col_idx"] if self.state_chunks
            else self.f0 + self.n_fields - 1)
        # Hidden injective row token: the row's opaque helper key, written as a
        # LITERAL in both workbook twins after the state chunks. Spot Check
        # re-derives the selected row's data-sheet rows by MATCHing this token
        # into each side's literal "Key (helper)" column, so its row matching
        # never trusts Comparison's own row links or status (CMP-AUD-218).
        self.c_token_idx = last_state_col_idx + 1
        self.c_token = get_column_letter(self.c_token_idx)
        self.c_token_header = f"__{_HELPER_KEY_VERSION}_TOKEN"
        self.comparison_physical_n_cols = self.c_token_idx
        self.comparison_physical_last_col = get_column_letter(
            self.comparison_physical_n_cols)

    def data_col(self, field_idx):
        """Data-sheet column letter for header[field_idx] (the data sheets
        carry a leading "Comparison row" link column)."""
        return get_column_letter(field_idx + 2 + self.off)

    def snapshot_sheet(self, side):
        try:
            return self.snapshot_sheets[side]
        except KeyError as exc:
            raise ValueError(f"unknown comparison side {side!r}") from exc

    def field_pos(self, field_idx):
        """0-based position of header[field_idx] among the displayed fields."""
        return self._field_pos[field_idx]

    def field_col(self, field_idx):
        """Comparison-sheet column letter for header[field_idx]."""
        return get_column_letter(self.f0 + self._field_pos[field_idx])

    def medwid_helper_col(self, field_idx, stage):
        """Data-sheet column for one versioned hidden Med-Wid helper stage."""
        return self._medwid_helper_cols[field_idx][stage]

    def medwid_canon_col(self, field_idx):
        return self.medwid_helper_col(field_idx, "CANON")

    def state_location(self, field_idx):
        """Return ``(chunk, one_based_character_offset)`` for one field."""
        return self._state_field_location[field_idx]

    def state_ref(self, row_ref, field_idx):
        """Formula expression projecting one field's code from its mask."""
        chunk, offset = self.state_location(field_idx)
        return f'MID(${chunk["col"]}{row_ref},{offset},1)'

    def key_expr(self, _r, helper_token=None):
        """Exact opaque lookup literal for a Comparison/Only-in formula.

        ``_r`` remains in the private signature for focused legacy callers;
        row identity is no longer reconstructed by delimiter concatenation.
        """
        if (type(helper_token) is not str
                or not helper_token.startswith(f"__{_HELPER_KEY_VERSION}_")):
            raise ValueError("a versioned opaque helper token is required")
        return _formula_text(helper_token)



# =============================================================================
# Workbook writing
# =============================================================================


def _formula_text(value):
    """Quote an exact string literal for an OOXML/Excel formula."""
    return '"' + value.replace('"', '""') + '"'


def _nested_substitutes(expr, replacements):
    for old, new in replacements:
        expr = f'SUBSTITUTE({expr},{_formula_text(old)},{_formula_text(new)})'
    return expr


def _checked_formula(formula, purpose):
    """Fail explicitly before save if a generated formula exceeds Excel's cap."""
    if len(formula) > _EXCEL_FORMULA_LIMIT:
        raise ValueError(
            f"{purpose} formula is {len(formula):,} characters; Excel allows "
            f"at most {_EXCEL_FORMULA_LIMIT:,}.")
    return formula


def _medwid_helper_formulas(source_ref, trimmed_ref, core_ref, valid_ref,
                            mask_ref):
    """Five proven legacy-scalar formulas for one Med-Wid source cell.

    Staging keeps every formula small, works in the installed non-dynamic Excel
    formula dialect, and gives Comparison/Spot Check a short CANON lookup. X/Y
    are internal sentinels only after VALID proves CORE contains ASCII digits
    and at most one dot, so they cannot collide with source data.
    """
    suffixes = _formula_text(_MEDWID_SUFFIXES)
    trimmed = f'=IF(ISBLANK({source_ref}),"",TRIM({source_ref}))'
    has_suffix = (
        f'AND({trimmed_ref}<>"",ISNUMBER(FIND(RIGHT({trimmed_ref},1),'
        f'{suffixes})))')
    core = (f'=IF({has_suffix},LEFT({trimmed_ref},LEN({trimmed_ref})-1),'
            f'{trimmed_ref})')

    dot_count = f'LEN({core_ref})-LEN(SUBSTITUTE({core_ref},".",""))'
    no_non_digits = _nested_substitutes(
        f'SUBSTITUTE({core_ref},".","")',
        ((str(i), "") for i in range(10)))
    valid = (f'=AND({core_ref}<>"",LEFT({core_ref},1)<>".",'
             f'RIGHT({core_ref},1)<>".",{dot_count}<=1,'
             f'{no_non_digits}="")')

    significance_mask = _nested_substitutes(
        f'SUBSTITUTE({core_ref},".","X")',
        ((str(i), "X") for i in range(1, 10)))
    mask = f'=IF({valid_ref},{significance_mask},"")'

    first = f'FIND("X",{mask_ref}&"X")'
    n_significant = f'LEN({mask_ref})-LEN(SUBSTITUTE({mask_ref},"X",""))'
    last = (f'IF({n_significant}=0,0,FIND("Y",SUBSTITUTE('
            f'{mask_ref},"X","Y",{n_significant})))')
    dot = f'FIND(".",{core_ref}&".")'
    plain = (f'IF({first}>LEN({core_ref}),"0",MID('
             f'{core_ref},{first},LEN({core_ref})))')
    raw_decimal = (f'IF({first}={dot},"0","")&MID('
                   f'{core_ref},{first},{last}-{first}+1)&"Y"')
    decimal = f'SUBSTITUTE(SUBSTITUTE({raw_decimal},".Y","Y"),"Y","")'
    suffix = (f'IF(LEN({core_ref})<LEN({trimmed_ref}),'
              f'RIGHT({trimmed_ref},1),"")')
    canonical = (f'=IF({trimmed_ref}="","",IF(NOT({valid_ref}),'
                 f'{trimmed_ref},IF({dot_count}=0,{plain},{decimal})&'
                 f'{suffix}))')

    formulas = (trimmed, core, valid, mask, canonical)
    return tuple(_checked_formula(formula, f"Med-Wid {stage}")
                 for stage, formula in zip(_MEDWID_HELPER_STAGES, formulas))


def _medwid_helper_values(value):
    """Literal twin of the five stages for the plain-values workbook."""
    trimmed = _xl_trim(value)
    has_suffix = bool(trimmed) and trimmed[-1] in _MEDWID_SUFFIXES
    core = trimmed[:-1] if has_suffix else trimmed
    valid = re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", core) is not None
    mask = ("".join("X" if ch == "." or ch in "123456789" else "0"
                    for ch in core) if valid else "")
    return trimmed, core, valid, mask, _medwid_norm(trimmed)


def _trim_ref(sheet, col, row_ref):
    """Blank-safe Excel twin of :func:`_xl_trim` for an indexed source cell.

    Excel ordinarily coerces ``INDEX`` of a truly blank cell to numeric zero.
    Testing the indexed reference with ``ISBLANK`` first preserves the policy's
    typed blank-vs-zero distinction before ``TRIM`` performs its ASCII-space
    normalization.
    """
    idx = f'INDEX({_sref(sheet)}!{col}:{col},{row_ref})'
    return f'IF(ISBLANK({idx}),"",TRIM({idx}))'


def _row_link(side, key, lay):
    """The '<side> Row' cell: the matched data-sheet row number as a CLICKABLE
    link, so a doubter can eyeball the source values instead of trusting the
    lookup. The link targets the ENTIRE ROW ("57:57"), so Excel selects the
    whole row on arrival — a temporary highlight that clears on the next
    click — while the view stays at the frozen left columns.
    (A bounded range like A57:AH57 made Excel scroll to the range's RIGHT
    edge when it didn't fit the window — measured via COM on real Excel;
    row-only references keep scrollColumn at home.) HYPERLINK's friendly
    value is the MATCH number itself, so the cell still counts as a number
    (the Summary SELF-CHECK relies on COUNT) — the MATCH is computed three
    times (range start, range end, display)."""
    s = _sref(side)
    m = f'MATCH({key},{s}!${lay.key_col}:${lay.key_col},0)'
    return f'=IFERROR(HYPERLINK("#{s}!"&{m}&":"&{m},{m}),"")'


def _row_link_value(side, row_num, lay):
    """_row_link with the row number known at build time (values workbook):
    same whole-row jump-and-select, no MATCH."""
    return f'=HYPERLINK("#{_sref(side)}!{row_num}:{row_num}",{row_num})'


def _link_font():
    return Font(name="Arial", size=10, color="0563C1", underline="single")


def _medwid_ref(lay, sheet, field_idx, row_ref):
    """Short lookup of the field's live hidden CANON helper."""
    col = lay.medwid_canon_col(field_idx)
    return f'INDEX({_sref(sheet)}!{col}:{col},{row_ref})'


def _isditto_xl(trim_ref):
    """Excel expression: TRUE when an already-TRIM'd cell ref is a ditto marker
    (non-empty and only '+' characters) — the formula twin of _is_plus_run."""
    return f'AND({trim_ref}<>"",SUBSTITUTE({trim_ref},"+","")="")'


def _matched_state_expr(sc, field_idx, eq, trim_t, trim_n):
    """Live formula expression for one matched compared-cell state code."""
    if sc.is_context(field_idx):
        return '"N"'
    if sc.ditto_nonasserting:
        return (f'IF(OR({_isditto_xl(trim_t)},{_isditto_xl(trim_n)}),"N",'
                f'IF({eq},"E","D"))')
    return f'IF({eq},"E","D")'


def _field_state_expr(lay, r, field_idx):
    """Matched-row state formula twin for one displayed field.

    Ordinary values compare the blank-safe ASCII-TRIM operands with ``EXACT``;
    Med-Wid compares the staged CANON helpers with ``EXACT``.  Context and
    Highway-Log ditto cells are explicit ``N`` rather than forged equality.
    """
    sc = lay.sc
    col = lay.data_col(field_idx)
    ct, cn = f"${lay.c_trow}{r}", f"${lay.c_nrow}{r}"
    trim_t = _trim_ref(sc.side_a, col, ct)
    trim_n = _trim_ref(sc.side_b, col, cn)
    if sc.is_medwid(field_idx):
        eq = (f'EXACT({_medwid_ref(lay, sc.side_a, field_idx, ct)},'
              f'{_medwid_ref(lay, sc.side_b, field_idx, cn)})')
    else:
        eq = f'EXACT({trim_t},{trim_n})'
    return _matched_state_expr(sc, field_idx, eq, trim_t, trim_n)


def _state_chunk_formula(lay, r, fields):
    """One live row-mask formula (unvalidated so the planner can measure it)."""
    matched = "&".join(_field_state_expr(lay, r, f) for f in fields)
    return (f'=IF(${lay.c_status}{r}<>"Both",REPT("U",{len(fields)}),'
            f'{matched})')


def _plan_state_chunk_fields(lay):
    """Greedily group fields into stable, Excel-safe state-mask chunks.

    Planning at ``XL_MAX_ROWS`` accounts for the longest possible row
    references.  A single field that cannot fit Excel's formula ceiling fails
    immediately and explicitly; no truncated or lossy fallback is permitted.
    """
    groups = []
    current = []
    for field_idx in lay.field_indices:
        single = _state_chunk_formula(lay, XL_MAX_ROWS, [field_idx])
        if len(single) > _EXCEL_FORMULA_LIMIT:
            raise ValueError(
                f"State-mask formula for field {field_idx} "
                f"({lay.sc.header[field_idx]!r}) is {len(single):,} characters; "
                f"Excel allows at most {_EXCEL_FORMULA_LIMIT:,}.")
        candidate = current + [field_idx]
        candidate_formula = _state_chunk_formula(lay, XL_MAX_ROWS, candidate)
        if current and len(candidate_formula) > _STATE_FORMULA_TARGET:
            groups.append(tuple(current))
            current = [field_idx]
        else:
            current = candidate
    if current:
        groups.append(tuple(current))
    for chunk_no, fields in enumerate(groups, start=1):
        _checked_formula(_state_chunk_formula(lay, XL_MAX_ROWS, fields),
                         f"state-mask chunk {chunk_no}")
    return tuple(groups)


def _mask_diff_count_expr(lay, r):
    """Count ``D`` codes across every hidden state chunk on one row."""
    if not lay.state_chunks:
        return "0"
    first = lay.state_chunks[0]["col"]
    last = lay.state_chunks[-1]["col"]
    refs = f'${first}{r}:${last}{r}'
    return (f'SUMPRODUCT(LEN({refs})-LEN(SUBSTITUTE({refs},"D","")))')


def _eq_with_ditto(sc, eq, trim_t, trim_n):
    """Wrap an equality expression so a ditto on EITHER side counts as equal
    (non-asserting), when the schema enables it. Otherwise returns `eq`
    unchanged, so non-Highway-Log comparisons stay byte-identical."""
    if sc.ditto_nonasserting:
        return f'OR({_isditto_xl(trim_t)},{_isditto_xl(trim_n)},{eq})'
    return eq


def _field_display_expr(lay, field_idx, status_ref, trim_t, trim_n, state_ref):
    """Formula expression projecting familiar display text from state truth."""
    if lay.sc.is_context(field_idx):
        matched = f'IF({trim_t}="",{trim_n},{trim_t})'
    else:
        show_t = f'IF({trim_t}="","(blank)",{trim_t})'
        show_n = f'IF({trim_n}="","(blank)",{trim_n})'
        matched = (f'IF({state_ref}="D",{show_t}&"{_DIFF_MARK}"&{show_n},'
                   f'{trim_t})')
    return (f'IF({status_ref}="{lay.only_a}",{trim_t},'
            f'IF({status_ref}="{lay.only_b}",{trim_n},{matched}))')


def _field_formula(lay, r, field_idx):
    """Comparison cell formula for data field `field_idx` (1-based into the
    header) on Comparison row `r`: the matched value when the two systems
    agree, 'a ≠ b' when they differ, and on single-side rows (tinted
    yellow/blue) that system's own value, so the row's data is still readable
    instead of blank. Excel's IF evaluates only the taken branch, so the
    absent side's INDEX (whose row ref is "") is never computed on
    single-side rows."""
    sc = lay.sc
    col = lay.data_col(field_idx)
    ct, cs = f"${lay.c_trow}{r}", f"${lay.c_nrow}{r}"
    t, n = _trim_ref(sc.side_a, col, ct), _trim_ref(sc.side_b, col, cs)
    st = f"${lay.c_status}{r}"
    # Production layouts always project from the hidden mask.  The fallback is
    # retained for small legacy unit fakes that exercise this private formula
    # helper without constructing a full _Layout; it still uses the same typed
    # state expression and never scans presentation text.
    state = (lay.state_ref(r, field_idx) if hasattr(lay, "state_ref")
             else _field_state_expr(lay, r, field_idx))
    return _checked_formula(
        "=" + _field_display_expr(lay, field_idx, st, t, n, state),
        f"Comparison display field {field_idx}")


def _field_value(sc, rt, rn, off, f, state_code=None):
    """What _field_formula DISPLAYS, computed in Python — the values
    workbook's cell for header[f]. `rt`/`rn` are the raw input rows
    (None when that side lacks the key); returns "" for an empty result.

    Writers always pass the retained state-mask code.  ``None`` remains a small
    compatibility seam for focused legacy checks that call this private helper
    directly; it derives the same typed state before projecting the display.
    """
    if state_code is None:
        state_code = ("U" if rt is None or rn is None
                      else compared_cell(sc, f, rt, rn, off).state_code)
    if rt is None:                       # B-only row: that side's own value
        if state_code != "U":
            raise ValueError(f"B-only field {f} has state {state_code!r}, not 'U'")
        return _xl_trim(rn[f + off])
    if rn is None:                       # A-only row
        if state_code != "U":
            raise ValueError(f"A-only field {f} has state {state_code!r}, not 'U'")
        return _xl_trim(rt[f + off])
    if state_code not in ("E", "D", "N"):
        raise ValueError(f"matched field {f} has invalid state {state_code!r}")
    trim_t, trim_n = _xl_trim(rt[f + off]), _xl_trim(rn[f + off])
    if sc.is_context(f):
        return trim_t if trim_t else trim_n
    if state_code == "D":
        return (f"{trim_t or '(blank)'}{_DIFF_MARK}"
                f"{trim_n or '(blank)'}")
    return trim_t


# The workbook is written in openpyxl's STREAMING (write_only) mode: the
# consolidated comparison carries ~2 million formula cells, which the normal
# in-memory mode cannot save in reasonable time or RAM (the consolidators use
# the same mode for the same reason). Streaming rules: sheets are created in
# display order, freeze/widths/filter/CF are set before rows are appended,
# and styled cells are WriteOnlyCells.

# Spreadsheet literal-cell guard. A free-text value beginning with one of
# these would be interpreted by Excel as a formula (the classic CSV/XLSX
# injection vector — a malicious Description like "=cmd|'/C calc'!A1" runs on
# open). openpyxl also classifies Excel's seven literal error tokens as live
# error cells unless their type is overridden. `guard=True` forces both classes
# to STRING cells so Excel shows source text verbatim and never executes or
# propagates it. The value is kept byte-for-byte (only the cell TYPE changes),
# so comparison equality and clean-data output are unaffected.
_FORMULA_LEAD = ("=", "+", "-", "@")
_EXCEL_ERROR_CODES = frozenset({
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A",
})


def is_formula_injection(value):
    """True when source text must be forced to an Excel string cell.

    The historical public name is retained for consolidator compatibility; it
    now also covers openpyxl's literal error tokens, which are just as unsafe to
    leave auto-typed at a source-data write boundary.
    """
    return (isinstance(value, str)
            and (value[:1] in _FORMULA_LEAD or value in _EXCEL_ERROR_CODES))


def _is_finite_source_numeric(value):
    """True for a supported finite non-Boolean source numeric."""
    if type(value) is bool:
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return math.isfinite(value)
    if isinstance(value, Decimal):
        return value.is_finite()
    return False


def _is_nonfinite_source_numeric(value):
    """True when Excel would silently lose a float/Decimal as a blank cell."""
    return ((isinstance(value, float) and not math.isfinite(value))
            or (isinstance(value, Decimal) and not value.is_finite()))


def _numeric_significant_digits(value):
    """Decimal significant digits needed to preserve a finite numeric."""
    if type(value) is bool or _is_nonfinite_source_numeric(value):
        return None
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = repr(value)
    elif isinstance(value, Decimal):
        text = str(value)
    else:
        return None
    try:
        digits = list(Decimal(text).as_tuple().digits)
    except (InvalidOperation, ValueError):  # silent-ok: unknown numeric subclass stays ordinary
        return None
    while len(digits) > 1 and digits[-1] == 0:
        digits.pop()
    return len(digits) or 1


def _excel_numeric_needs_text(value):
    digits = _numeric_significant_digits(value)
    return digits is not None and digits > 15


def set_safe_literal_cell(cell, value, *, exact_source_numeric=False):
    """Assign an exact source literal without inviting Excel coercion.

    Formula-leading/error-looking strings stay text as before.  Actual Boolean
    values are emitted as uppercase text so Python and Excel see the same
    case-sensitive operand; numeric 1/0 and int subclasses are not Boolean.
    With ``exact_source_numeric=True``, every finite source numeric is emitted
    as the exact text used by ``_xl_trim``.
    This is intentionally stricter than a 15-digit check: Excel's implicit text
    conversion can change Decimal scale (``1.2300``), exponent spelling
    (``1E+3``), or finite float notation even when precision itself is safe.
    Without that source-only flag, the shared legacy guard still protects any
    numeric requiring more than Excel's 15 significant digits while leaving
    small engine-owned counts/occurrences numeric. Non-finite float/Decimal
    values are rejected rather than serialized by openpyxl as deceptive blank
    cells. The helper works for normal and WriteOnly cells.
    """
    if type(value) is bool:
        cell.value = "TRUE" if value else "FALSE"
        cell.data_type = "s"
        cell.number_format = "@"
    elif _is_nonfinite_source_numeric(value):
        raise ValueError("non-finite numeric source values cannot be compared")
    elif ((exact_source_numeric and _is_finite_source_numeric(value))
          or _excel_numeric_needs_text(value)):
        cell.value = _xl_trim(value)
        cell.data_type = "s"
        cell.number_format = "@"
    else:
        cell.value = value
    if is_formula_injection(cell.value):
        cell.data_type = "s"
    return cell


def _styled(ws, value, font, fill=None, align=None, guard=False,
            exact_source_numeric=False):
    c = WriteOnlyCell(ws, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if guard:
        set_safe_literal_cell(
            c, value, exact_source_numeric=exact_source_numeric)
    return c


def _header_row(ws, values, comment_fn=None):
    font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color=_DARK)
    align = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    cells = [_styled(ws, v, font, fill, align) for v in values]
    if comment_fn is not None:                # attach hover tooltips (HL columns)
        for c in cells:
            cm = comment_fn(c.value)
            if cm is not None:
                c.comment = cm
    return cells


def _apply_field_widths(ws, widths, col_for, lay):
    """Set column widths from a {field name -> width} schema dict. Applies to
    EVERY column with that name (header.index would width only the first of a
    duplicate name); skips a column only when col_for can't place it — the
    Comparison/Only-in field_col raises for the key index, while the data
    sheet's data_col places the key column fine (so its width is kept)."""
    for name, width in widths.items():
        for idx, h in enumerate(lay.sc.header):
            if h != name:
                continue
            try:
                col = col_for(idx)
            except (KeyError, ValueError):
                continue
            ws.column_dimensions[col].width = width


def _write_data_sheet(wb, name, tab_color, rows, lay, events, cmp_rows,
                      helper_keys=None, live_medwid_helpers=True):
    """One input copied to its sheet, with a leading 'Comparison row' LINK
    back to where each row appears on the Comparison sheet (column A — so a
    reviewer who jumped here from a row link has a one-click way back, right
    where they land) and the literal 'Key (helper)' column after the data. The link
    target is a literal (cmp_rows[i]), consistent with the workbook's
    design: the comparison's row universe is fixed at build time, only the
    VALUES are live. Versioned hidden Med-Wid helpers follow Key(helper) when
    the schema needs them; visible/filter/freeze geometry is unchanged. The
    formulas flavor writes live stages; the plain-values flavor writes their
    exact Python literals. Versioned build-freshness chunks then prove every
    source/helper cell still equals its very-hidden build snapshot."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    body_font = Font(name="Arial", size=10)
    link_font = _link_font()

    # Keep the back-link + Route + key column in view while scrolling fields.
    ws.freeze_panes = "D2" if lay.has_route else "C2"
    ws.auto_filter.ref = f"A1:{lay.data_last_col}{len(rows) + 1}"
    ws.column_dimensions[lay.key_col].width = 14
    ws.column_dimensions[lay.back_col].width = 13
    for f in lay.medwid_field_indices:
        for stage in _MEDWID_HELPER_STAGES:
            ws.column_dimensions[lay.medwid_helper_col(f, stage)].hidden = True
    for column in lay.build_fresh_cols:
        ws.column_dimensions[column].hidden = True
    if lay.has_route:
        ws.column_dimensions[lay.route_data_col].width = 8
    _apply_field_widths(ws, lay.sc.data_widths, lay.data_col, lay)

    ws.append(_header_row(
        ws,
        (["Comparison row"] + lay.data_header + ["Key (helper)"]
         + lay.medwid_helper_headers + lay.build_fresh_headers),
        lay.sc.header_comment))
    # DISPLAY-only ditto resolution (Highway Log): {row_index: {col_in_row:
    # resolved}} — keeps the raw `++` in the cell (the non-asserting diff needs
    # it) but tints it and shows the paired-roadbed value on hover. None elsewhere.
    fills = (lay.sc.ditto_resolver(rows, lay.has_route)
             if lay.sc.ditto_resolver is not None else None)
    ditto_fill = PatternFill("solid", start_color=_DITTO_FILL) if fills else None
    for r, row in enumerate(rows, start=2):
        u = cmp_rows[r - 2]
        # Whole-row target: the jump selects the entire Comparison row
        # (temporary highlight until the next click) WITHOUT scrolling
        # right, same as the forward row links.
        cells = [_styled(ws, f'=HYPERLINK("#Comparison!{u}:{u}",{u})', link_font)]
        # Raw input values are guarded: a Description like "=cmd…" stays text,
        # never a live formula (the field FORMULAS read these via TRIM/INDEX, so
        # guarding here protects both flavors at the source).
        cells += [_styled(ws, v, body_font, guard=True,
                          exact_source_numeric=True) for v in row]
        row_fills = fills.get(r - 2) if fills else None
        if row_fills:                        # mark each dittoed cell in this row
            for col_in_row, resolved in row_fills.items():
                c = cells[1 + col_in_row]    # cells[0] is the back-link
                c.fill = ditto_fill
                shown = resolved if resolved not in (None, "") else "(no paired value found)"
                c.comment = Comment(
                    f"Ditto ('{c.value}') — value is on the paired roadbed's row; "
                    f"resolves to: {shown}. Not counted as a difference.",
                    "TSMIS Exporter")
        # Both flavors write a LITERAL key (the comparison's row universe is
        # build-time static): a live COUNTIFS mis-numbers blank key fields, and
        # the literal always matches the Comparison sheet's stored occurrence #.
        # Guard the helper key too: a key like "=X" would otherwise become a
        # live formula here, breaking the Comparison MATCH lookups (which search
        # for the literal "=X|1") AND splitting the two flavors. Guarding stores
        # the literal string the lookups expect.
        cells.append(_styled(ws, helper_keys[r - 2], body_font, guard=True))
        # E1 formula twin: five short, hidden stages per Med-Wid field. Physical
        # geometry is identical between flavors; formula mode stays live, while
        # values mode stores exact literals and therefore remains fast/plain.
        for f in lay.medwid_field_indices:
            if live_medwid_helpers:
                cols = [lay.medwid_helper_col(f, stage)
                        for stage in _MEDWID_HELPER_STAGES]
                refs = [f"${col}{r}" for col in cols]
                source_ref = f"${lay.data_col(f)}{r}"
                helpers = _medwid_helper_formulas(
                    source_ref, refs[0], refs[1], refs[2], refs[3])
                cells.extend(_styled(ws, helper, body_font)
                             for helper in helpers)
            else:
                source_value = row[f + lay.off]
                helpers = _medwid_helper_values(source_value)
                cells.extend(_styled(ws, helper, body_font,
                                     guard=isinstance(helper, str))
                             for helper in helpers)
        for chunk_no, columns in enumerate(lay.build_fresh_chunks, start=1):
            cells.append(_styled(
                ws,
                _checked_formula(
                    _build_fresh_formula(lay, name, r, columns),
                    f"{name} build-snapshot chunk {chunk_no}"),
                body_font))
        ws.append(cells)
        if (r - 1) % _PROGRESS_EVERY == 0:
            events.on_log(f"  {name} sheet: {r - 1:,} rows…")
            if events.is_cancelled():
                return None
    # A hidden footer makes below-table appends observable without scanning
    # whole columns in every Summary recalculation. It sits outside the visible
    # filter/data range; each chunk scans only its source/helper columns below
    # the original row universe.
    footer_row = len(rows) + 2
    footer = [None] * lay.snapshot_physical_n_cols
    for chunk_no, columns in enumerate(lay.build_fresh_chunks, start=1):
        footer.append(_styled(
            ws,
            _checked_formula(
                _build_tail_formula(footer_row, columns),
                f"{name} appended-row sentinel chunk {chunk_no}"),
            body_font))
    ws.append(footer)
    return ws


def _write_snapshot_sheet(wb, side, rows, lay, events, helper_keys):
    """Very-hidden immutable build snapshot for one data sheet.

    It mirrors columns A through the last Med-Wid helper exactly, replacing the
    visible back-link with a stable source-row ordinal. Data-sheet freshness
    formulas compare B:last-helper against the same row here. Keeping the copy
    on a separate sheet makes insert/delete/reorder operations detectable too.
    """
    name = lay.snapshot_sheet(side)
    ws = wb.create_sheet(name)
    ws.sheet_state = "veryHidden"
    body_font = Font(name="Arial", size=10)
    ws.append(_header_row(
        ws,
        (["Source row"] + lay.data_header + ["Key (helper)"]
         + lay.medwid_helper_headers)))
    if len(helper_keys) != len(rows):
        raise ValueError("snapshot helper-key count does not match source rows")
    for row_number, row in enumerate(rows, start=2):
        cells = [_styled(ws, row_number - 1, body_font)]
        cells.extend(_styled(
            ws, value, body_font, guard=True, exact_source_numeric=True)
            for value in row)
        cells.append(_styled(
            ws, helper_keys[row_number - 2], body_font, guard=True))
        for field_idx in lay.medwid_field_indices:
            source_value = row[field_idx + lay.off]
            cells.extend(_styled(
                ws, helper, body_font, guard=isinstance(helper, str))
                for helper in _medwid_helper_values(source_value))
        ws.append(cells)
        if (row_number - 1) % _PROGRESS_EVERY == 0:
            events.on_log(
                f"  Build snapshot for {side}: {row_number - 1:,} rows…")
            if events.is_cancelled():
                return None
    return ws


def _write_comparison(wb, union, lay, events, vals=None, helper_tokens=None):
    """The big sheet. `vals` None = live formulas (the default workbook);
    else the values model (run_compare builds it) and every cell is the
    computed RESULT — identical text, no formulas, links kept."""
    sc = lay.sc
    ws = wb.create_sheet("Comparison")
    ws.sheet_properties.tabColor = _TAB_COLORS["comparison"]
    body_font = Font(name="Arial", size=10)

    last = len(union) + 1
    ws.row_dimensions[1].height = 45.75
    ws.freeze_panes = f"{lay.first_field_col}2"
    ws.auto_filter.ref = f"A1:{lay.last_field_col}{last}"
    if lay.has_route:
        ws.column_dimensions["A"].width = 8
    ws.column_dimensions[lay.c_loc].width = 12
    ws.column_dimensions[lay.c_occ].width = 4
    ws.column_dimensions[lay.c_trow].width = 7
    ws.column_dimensions[lay.c_status].width = 11
    ws.column_dimensions[lay.c_diffs].width = 6
    _apply_field_widths(ws, sc.cmp_widths, lay.field_col, lay)

    for chunk in lay.state_chunks:
        ws.column_dimensions[chunk["col"]].hidden = True
    ws.column_dimensions[lay.c_token].hidden = True

    # Conditional formatting keeps the familiar presentation but reads the
    # hidden state code, never display text.  One rule per state chunk lets the
    # relative field position project directly into that chunk's mask.
    f1 = lay.first_field_col
    full = f"A2:{lay.last_field_col}{last}"
    for chunk in lay.state_chunks:
        first_field = lay.field_col(chunk["fields"][0])
        last_field = lay.field_col(chunk["fields"][-1])
        fields = f"{first_field}2:{last_field}{last}"
        ws.conditional_formatting.add(fields, FormulaRule(
            formula=[f'MID(${chunk["col"]}2,COLUMN()-'
                     f'COLUMN(${first_field}$2)+1,1)="D"'],
            fill=PatternFill(bgColor="FFC7CE"),
            font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add(full, FormulaRule(
        formula=[f'${lay.c_status}2="{lay.only_a}"'], fill=PatternFill(bgColor="FFE699")))
    ws.conditional_formatting.add(full, FormulaRule(
        formula=[f'${lay.c_status}2="{lay.only_b}"'], fill=PatternFill(bgColor="BDD7EE")))
    ws.conditional_formatting.add(f"{lay.c_diffs}2:{lay.c_diffs}{last}", CellIsRule(
        operator="greaterThan", formula=["0"],
        font=Font(color="C00000", bold=True)))

    link_font = _link_font()
    link_cols = {3, 4} if lay.has_route else {2, 3}   # trow / nrow positions
    ws.append(_header_row(ws, lay.id_headers
                          + [sc.header[i] for i in lay.field_indices]
                          + [chunk["header"] for chunk in lay.state_chunks]
                          + [lay.c_token_header],
                          lay.sc.header_comment))
    if not isinstance(helper_tokens, dict) or set(helper_tokens) != set(union):
        raise ValueError(
            "Comparison helper-token map must cover the union exactly")
    if len(set(helper_tokens.values())) != len(union):
        raise ValueError("Comparison helper tokens must be injective")
    for i, (route, loc, occ) in enumerate(union):
        r = i + 2
        visible_loc = _visible_key(loc)
        if vals is None:
            key = _formula_text(helper_tokens[(route, loc, occ)])
            status_formula = (
                f'=IF(AND({lay.c_trow}{r}<>"",{lay.c_nrow}{r}<>""),"Both",'
                f'IF({lay.c_trow}{r}<>"","{lay.only_a}","{lay.only_b}"))')
            state_formulas = [
                _checked_formula(_state_chunk_formula(lay, r, chunk["fields"]),
                                 f"Comparison state-mask chunk {chunk_no}")
                for chunk_no, chunk in enumerate(lay.state_chunks, start=1)
            ]
            diffs_formula = _checked_formula(
                f'=IF({lay.c_status}{r}<>"Both","",'
                f'{_mask_diff_count_expr(lay, r)})',
                "Comparison Diffs")
            row = ([route] if lay.has_route else []) + [
                visible_loc, occ,
                _row_link(sc.side_a, key, lay),
                _row_link(sc.side_b, key, lay),
                status_formula,
                diffs_formula,
            ]
            row += [_field_formula(lay, r, f)
                    for f in lay.field_indices]
            row += state_formulas
            row.append(helper_tokens[(route, loc, occ)])
        else:
            k = (route, loc, occ)
            rt, rn = vals["by_t"].get(k), vals["by_n"].get(k)
            tr, nr = vals["row_t"].get(k), vals["row_n"].get(k)
            status = ("Both" if rt is not None and rn is not None
                      else lay.only_a if rt is not None else lay.only_b)
            mask = vals["state_masks"][k]
            if len(mask) != lay.n_fields:
                raise ValueError(
                    f"state mask for {k!r} has {len(mask)} codes; "
                    f"expected {lay.n_fields}")
            fields = []
            for pos, f in enumerate(lay.field_indices):
                v = _field_value(sc, rt, rn, vals["off"], f, mask[pos])
                fields.append(v if v != "" else None)
            mask_chunks = [mask[chunk["start"]:chunk["end"] + 1]
                           for chunk in lay.state_chunks]
            ndiff = mask.count("D")
            row = ([route] if lay.has_route else []) + [
                visible_loc, occ,
                _row_link_value(sc.side_a, tr, lay) if tr else None,
                _row_link_value(sc.side_b, nr, lay) if nr else None,
                status,
                ndiff if status == "Both" else None,
            ] + fields + mask_chunks + [helper_tokens[k]]
        # Guard literal cells against formula injection without touching our own
        # formulas: in the formulas flavor only the key/route id cells and the
        # trailing row token are literals (everything else is an =formula or a
        # HYPERLINK link); in the values flavor every non-link cell is a
        # literal. (HYPERLINK cells must stay formulas, so they're never
        # guarded.)
        if vals is None:
            guard_set = {0} if lay.has_route else set()
            guard_set.add(1 if lay.has_route else 0)         # the key (loc) cell
            guard_set.add(len(row) - 1)                      # the row token
        else:
            guard_set = set(range(len(row))) - link_cols
        ws.append([_styled(ws, v, link_font if j in link_cols else body_font,
                           guard=(j in guard_set))
                   for j, v in enumerate(row)])
        if (i + 1) % _PROGRESS_EVERY == 0:
            events.on_log(f"  Comparison sheet: {i + 1:,} of {len(union):,} rows…")
            if events.is_cancelled():
                return None
    return ws


def _write_spot_check(wb, lay, n_union, default_row, default_key,
                      manual_calc=False):
    """'Spot Check': one row under a microscope, for reviewers who doubt
    the formulas. Type any Comparison row number (or find one by its key)
    and the sheet lays that row out field by field — the RAW values from
    both data sheets next to an INDEPENDENTLY recomputed verdict (same
    TRIM / Med Wid rules, computed straight from the data sheets, never
    reading the Comparison sheet's answer) and an Agree? column that flags
    any disagreement with what the Comparison sheet displays. The two
    data-sheet rows themselves are derived INDEPENDENTLY too (CMP-AUD-218):
    the row's hidden key token (Comparison's trailing literal column) is
    MATCHed into each side's literal "Key (helper)" column, so a forged
    Comparison row link or status can never pick which rows get audited;
    the "Row integrity" line EXACT-compares Comparison's stored links and
    status against that independent derivation and shouts CHECK on any
    disagreement. On one-sided rows the verdict column carries the
    independently derived status itself (tinted) plus a loud callout line —
    the Status cell alone is easy to miss — and Agree? still verifies the
    displayed value against that system's data sheet. Opens pre-set to
    `default_row`, the first matched row with differences."""
    sc = lay.sc
    a, b = sc.side_a, sc.side_b
    ws = wb.create_sheet("Spot Check")
    ws.sheet_properties.tabColor = _TAB_COLORS["spot"]
    ws.sheet_view.showGridLines = False

    title_font = Font(name="Arial", size=14, bold=True, color=_DARK)
    note_font = Font(name="Arial", size=10, color="595959")
    banner_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    banner_fill = PatternFill("solid", start_color=_DARK)
    body_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    alert_font = Font(name="Arial", size=11, bold=True, color="C00000")
    input_font = Font(name="Arial", size=11, bold=True)
    input_fill = PatternFill("solid", start_color="FFF2CC")
    link_font = _link_font()
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    last = n_union + 1
    inp = "$C$6"                                   # the row-number input cell
    trow_cell, nrow_cell = "$C$12", "$F$12"        # matched data-sheet rows
    status = "$C$11"                               # Comparison's claimed status
    # Hidden independent-derivation cells (columns K:M are hidden anyway):
    # the selected row's key token, then its MATCHed row in each data sheet.
    token_cell = "$M$12"
    trow_ind, nrow_ind = "$K$12", "$L$12"
    # Banner and header occupy DISTINCT rows (CMP-AUD-214): the field-by-field
    # banner keeps row 15, the header row sits at F_FIRST - 1 = 16, and the
    # field rows start at 17 — the header no longer overwrites the banner.
    F_FIRST = 17                                   # first field row
    F_LAST = F_FIRST + lay.n_fields - 1
    has_medwid = bool(lay.medwid_field_indices)

    for col, w in (("A", 2), ("B", 19), ("C", 24), ("D", 24), ("E", 17),
                   ("F", 30), ("G", 9), ("H", 16), ("I", 16), ("J", 16)):
        ws.column_dimensions[col].width = w
    # Values workbooks keep their large data-sheet helpers literal. A tiny live
    # staged block here (ten cells per Med-Wid field row) preserves Spot Check's
    # genuinely independent recomputation after a reviewer edits source values.
    spot_med_a_cols = list(range(20, 25))   # T:X, CANON=X
    spot_med_b_cols = list(range(25, 30))   # Y:AC, CANON=AC
    for col_idx in range(11, 14):           # K:M, generic state/display twins
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    if has_medwid:
        for col_idx in spot_med_a_cols + spot_med_b_cols:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    # Verdict / agreement colors. One-sided verdicts tint like the
    # Comparison sheet's yellow/blue rows so the situation is unmissable.
    ws.conditional_formatting.add(f"E{F_FIRST}:E{F_LAST}", CellIsRule(
        operator="equal", formula=['"DIFFERENT"'],
        font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add(f"E{F_FIRST}:E{F_LAST}", CellIsRule(
        operator="equal", formula=[f'"{lay.only_a}"'],
        fill=PatternFill(bgColor="FFE699")))
    ws.conditional_formatting.add(f"E{F_FIRST}:E{F_LAST}", CellIsRule(
        operator="equal", formula=[f'"{lay.only_b}"'],
        fill=PatternFill(bgColor="BDD7EE")))
    ws.conditional_formatting.add(f"G{F_FIRST}:G{F_LAST}", CellIsRule(
        operator="equal", formula=['"CHECK"'],
        fill=PatternFill(bgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add(f"G{F_FIRST}:G{F_LAST}", CellIsRule(
        operator="equal", formula=['"OK"'], font=Font(color="2E7D32", bold=True)))
    # The Row-integrity verdict gets the same loud treatment.
    ws.conditional_formatting.add("C14", CellIsRule(
        operator="equal", formula=['"CHECK"'],
        fill=PatternFill(bgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add("C14", CellIsRule(
        operator="equal", formula=['"OK"'], font=Font(color="2E7D32", bold=True)))

    grid = {}

    def put(rc, value, font=body_font, fill=None, align=None, fmt=None):
        grid[rc] = (value, font, fill, align, fmt)

    def banner(row, text):
        put((row, 2), text, banner_font, banner_fill)
        for c in range(3, 11):                     # extend the band to col J
            put((row, c), "", banner_font, banner_fill)

    # --- intro + inputs --------------------------------------------------
    put((2, 2), f"Spot Check — audit any single {sc.id_noun}", title_font)
    put((3, 2), "Every value below recomputes for the row you pick. The "
                "'Independent verdict' column re-compares the two data sheets "
                "directly (TRIM" + (" + the Med Wid rule" if has_medwid else "")
                + ") WITHOUT reading the "
                "Comparison sheet — Agree? = OK means both computations "
                "reached the same answer. The two source rows themselves are "
                "key-matched independently of Comparison's row links; the "
                "'Row integrity' line flags any disagreement with what "
                "Comparison stored.", note_font)
    put((4, 2), f"In difference cells the order is always:   "
                f"{a} value{_DIFF_MARK}{b} value   ({a} first, {b} second).",
        bold_font)
    if manual_calc:
        put((5, 2), "▶ PRESS F9 AFTER EVERY CHANGE — this workbook calculates "
                    "manually, so nothing updates until you do.", alert_font)
    put((6, 2), f"Comparison row # to check (2–{last}):", bold_font)
    put((6, 3), default_row, input_font, input_fill, center)
    put((6, 4), "← type a row number" + (", then press F9" if manual_calc
                                         else " (updates instantly)"), note_font)
    d_route, d_loc, d_occ = default_key
    d_loc = _visible_key(d_loc)
    key_label = sc.header[lay.key_field]
    if lay.has_route:
        put((7, 2), "…or find one:", note_font)
        put((7, 3), "Route:", bold_font, None, right)
        put((7, 4), d_route, body_font, input_fill, center, "@")
        put((7, 5), f"{key_label}:", bold_font, None, right)
        put((7, 6), d_loc, body_font, input_fill, center, "@")
        put((7, 7), "Occ #:", bold_font, None, right)
        put((7, 8), d_occ, body_font, input_fill, center)
        find = (f"SUMPRODUCT((Comparison!$A$2:$A${last}=$D$7)"
                f"*(Comparison!${lay.c_loc}$2:${lay.c_loc}${last}=$F$7)"
                f"*(Comparison!${lay.c_occ}$2:${lay.c_occ}${last}=$H$7)"
                f"*ROW(Comparison!$A$2:$A${last}))")
        put((7, 9), "→ Comparison row:", bold_font, None, right)
        put((7, 10), f'=IF({find}=0,"not found",{find})', bold_font, None, center)
    else:
        put((7, 2), "…or find one:", note_font)
        put((7, 3), f"{key_label}:", bold_font, None, right)
        put((7, 4), d_loc, body_font, input_fill, center, "@")
        put((7, 5), "Occ #:", bold_font, None, right)
        put((7, 6), d_occ, body_font, input_fill, center)
        find = (f"SUMPRODUCT((Comparison!$A$2:$A${last}=$D$7)"
                f"*(Comparison!$B$2:$B${last}=$F$7)"
                f"*ROW(Comparison!$A$2:$A${last}))")
        put((7, 7), "→ Comparison row:", bold_font, None, right)
        put((7, 8), f'=IF({find}=0,"not found",{find})', bold_font, None, center)

    # --- what the Comparison sheet says ----------------------------------
    def cmp_idx(col):
        return f"INDEX(Comparison!${col}:${col},{inp})"

    banner(9, "THE SELECTED ROW — COMPARISON'S CLAIMS + THE INDEPENDENT "
              "KEY MATCH")
    if lay.has_route:
        put((10, 2), "Route:", bold_font)
        put((10, 3), f'=IFERROR({cmp_idx(lay.c_route)},"")', body_font)
    put((10, 5), f"{key_label}:", bold_font)
    put((10, 6), f'=IFERROR({cmp_idx(lay.c_loc)},"")', body_font)
    put((10, 8), "Occurrence #:", bold_font)
    put((10, 9), f'=IFERROR({cmp_idx(lay.c_occ)},"")', body_font)
    put((11, 2), "Status:", bold_font)
    put((11, 3), f'=IFERROR({cmp_idx(lay.c_status)},"")', bold_font)
    put((11, 5), "Diffs counted:", bold_font)
    put((11, 6), f'=IFERROR({cmp_idx(lay.c_diffs)},"")', bold_font)
    # The two data-sheet rows are derived INDEPENDENTLY (CMP-AUD-218): the
    # row's hidden literal key token (Comparison's trailing token column) is
    # MATCHed into each side's literal "Key (helper)" column. Comparison's
    # own stored row links are never read here — they are only EXACT-compared
    # against this derivation on the Row-integrity line below, so a forged
    # link or status can no longer choose which source rows get audited.
    tok_col = lay.c_token
    tok_idx = f'INDEX(Comparison!${tok_col}:${tok_col},{inp})'
    put((12, 13), _checked_formula(                      # M12: the key token
        f'=IFERROR(IF({tok_idx}="","",{tok_idx}),"")',
        "Spot key token"), body_font)
    for column, side in ((11, a), (12, b)):              # K12/L12: MATCHed rows
        put((12, column), _checked_formula(
            f'=IF({token_cell}="","",IFERROR(MATCH({token_cell},'
            f'{_sref(side)}!${lay.key_col}:${lay.key_col},0),""))',
            f"Spot independent {side} row"), body_font)
    put((12, 2), f"{a} row (key-matched):", bold_font)
    put((12, 3), f'=IF({trow_ind}="","",'
                 f'HYPERLINK("#{_sref(a)}!"&{trow_ind}&'
                 f'":"&{trow_ind},{trow_ind}))', link_font)
    put((12, 5), f"{b} row (key-matched):", bold_font)
    put((12, 6), f'=IF({nrow_ind}="","",'
                 f'HYPERLINK("#{_sref(b)}!"&{nrow_ind}&'
                 f'":"&{nrow_ind},{nrow_ind}))', link_font)
    put((12, 8), "← matched from the row's hidden key token, never from "
                 "Comparison's stored links", note_font)
    # Loud one-sided callout, riding the INDEPENDENT membership: the Status
    # cell alone is easy to miss, and a forged status must not steer this.
    put((13, 2), f'=IF(AND({trow_ind}<>"",{nrow_ind}=""),'
                 f'"⚠ THIS {sc.id_noun.upper()} EXISTS ONLY IN '
                 f'{a} — there is no {b} row to compare; {b} values below '
                 f'are blank.",IF(AND({trow_ind}="",{nrow_ind}<>""),'
                 f'"⚠ THIS {sc.id_noun.upper()} EXISTS '
                 f'ONLY IN {b} — there is no {a} row to compare; {a} '
                 f'values below are blank.",""))', alert_font)
    # Row integrity: Comparison's stored links and status must EXACTLY equal
    # the independent derivation. A consistently relinked pair or a falsely
    # one-sided status/link set says CHECK here even when every downstream
    # display was recomputed to match the forgery.
    independent_row_status = (
        f'IF(AND({trow_ind}="",{nrow_ind}=""),"(no data-sheet match)",'
        f'IF({nrow_ind}="","{lay.only_a}",'
        f'IF({trow_ind}="","{lay.only_b}","Both")))')
    claimed_t, claimed_n = cmp_idx(lay.c_trow), cmp_idx(lay.c_nrow)
    put((14, 2), "Row integrity:", bold_font)
    put((14, 3), _checked_formula(
        f'=IF(OR({inp}<2,{token_cell}=""),"",'
        f'IF(AND(IF({claimed_t}="","",{claimed_t})={trow_ind},'
        f'IF({claimed_n}="","",{claimed_n})={nrow_ind},'
        f'EXACT({cmp_idx(lay.c_status)},{independent_row_status})),'
        f'"OK","CHECK"))',
        "Spot row integrity"), bold_font, None, center)
    put((14, 4), '=IF($C$14="CHECK","⚠ Comparison\'s stored row links or '
                 'status DISAGREE with the independent key match — do not '
                 'trust this row until the workbook is regenerated.","")',
        alert_font)

    # --- field-by-field ---------------------------------------------------
    banner(15, "FIELD BY FIELD — RECOMPUTED FROM THE DATA SHEETS "
               "(independent of the Comparison sheet)")
    headers = ["Field", f"{a} value (as stored)", f"{b} value (as stored)",
               "Independent verdict", f"Comparison sheet shows "
               f"({a}{_DIFF_MARK}{b})", "Agree?"]
    if has_medwid:
        headers += [f"{a} Med-Wid normalized", f"{b} Med-Wid normalized"]
    for j, h in enumerate(headers):
        put((F_FIRST - 1, 2 + j), h,
            Font(name="Arial", size=10, bold=True, color="FFFFFF"),
            banner_fill, Alignment(horizontal="center", wrap_text=True))
    if has_medwid:
        for side_token, columns in (("A", spot_med_a_cols),
                                    ("B", spot_med_b_cols)):
            for stage, column in zip(_MEDWID_HELPER_STAGES, columns):
                put((F_FIRST - 1, column),
                    f"__{_MEDWID_HELPER_VERSION}_SPOT_{side_token}_{stage}",
                    body_font)
    for column, label in (
            (11, f"__{_STATE_MASK_VERSION}_SPOT_INDEPENDENT_STATE"),
            (12, f"__{_STATE_MASK_VERSION}_SPOT_EXPECTED_DISPLAY"),
            (13, f"__{_STATE_MASK_VERSION}_SPOT_COMPARISON_STATE")):
        put((F_FIRST - 1, column), label, body_font)

    def raw(side, col, row_ref):
        idx = f"INDEX({_sref(side)}!{col}:{col},{row_ref})"
        return (f'=IF({row_ref}="","",IFERROR(IF(ISBLANK({idx}),"",{idx}),""))')

    for pos, f in enumerate(lay.field_indices):
        r = F_FIRST + pos
        col = lay.data_col(f)
        fcol = lay.field_col(f)
        is_date = sc.header[f] in sc.date_fields
        fmt = "mm/dd/yyyy" if is_date else None
        trim_t = _trim_ref(a, col, trow_cell)
        trim_n = _trim_ref(b, col, nrow_cell)
        if sc.is_medwid(f):
            a_letters = [get_column_letter(c) for c in spot_med_a_cols]
            b_letters = [get_column_letter(c) for c in spot_med_b_cols]
            a_refs = [f"${letter}{r}" for letter in a_letters]
            b_refs = [f"${letter}{r}" for letter in b_letters]
            a_idx = f'INDEX({_sref(a)}!{col}:{col},{trow_cell})'
            b_idx = f'INDEX({_sref(b)}!{col}:{col},{nrow_cell})'
            # INDEX of a truly blank source cell evaluates to numeric zero.
            # Preserve typed blank before the staged Med-Wid helper sees it;
            # otherwise Spot Check would canonically compare blank as "0".
            a_source = (f'IF({trow_cell}="","",'
                        f'IF(ISBLANK({a_idx}),"",{a_idx}))')
            b_source = (f'IF({nrow_cell}="","",'
                        f'IF(ISBLANK({b_idx}),"",{b_idx}))')
            a_helpers = _medwid_helper_formulas(
                a_source, a_refs[0], a_refs[1], a_refs[2], a_refs[3])
            b_helpers = _medwid_helper_formulas(
                b_source, b_refs[0], b_refs[1], b_refs[2], b_refs[3])
            for column, formula in zip(spot_med_a_cols, a_helpers):
                put((r, column), formula, body_font)
            for column, formula in zip(spot_med_b_cols, b_helpers):
                put((r, column), formula, body_font)
            a_canon, b_canon = a_refs[4], b_refs[4]
            eq = f'EXACT({a_canon},{b_canon})'
            put((r, 8), f'=IF({trow_cell}="","",'
                        f'{a_canon})', body_font,
                None, center)
            put((r, 9), f'=IF({nrow_cell}="","",'
                        f'{b_canon})', body_font,
                None, center)
        else:
            eq = f'EXACT({trim_t},{trim_n})'
        matched_state = _matched_state_expr(sc, f, eq, trim_t, trim_n)
        independent_status = (f'IF({trow_cell}="","{lay.only_b}",'
                              f'IF({nrow_cell}="","{lay.only_a}","Both"))')
        independent_state = (f'=IF({status}="","",IF(AND('
                             f'{trow_cell}<>"",{nrow_cell}<>""),'
                             f'{matched_state},"U"))')
        expected_display = _field_display_expr(
            lay, f, independent_status, trim_t, trim_n, f"$K{r}")
        chunk, state_offset = lay.state_location(f)
        comparison_mask = (f'INDEX(Comparison!${chunk["col"]}:'
                           f'${chunk["col"]},{inp})')
        comparison_display = f'INDEX(Comparison!{fcol}:{fcol},{inp})'
        put((r, 2), sc.header[f], bold_font)
        put((r, 3), raw(a, col, trow_cell), body_font, None, None, fmt)
        put((r, 4), raw(b, col, nrow_cell), body_font, None, None, fmt)
        # K:L independently recompute the typed state and complete familiar
        # display from raw source cells. M projects Comparison's hidden state.
        # The visible verdict is presentation only; Agree? EXACT-compares both
        # state and display, so literal separator content cannot forge truth.
        put((r, 11), _checked_formula(independent_state,
                                      f"Spot state field {f}"), body_font)
        put((r, 12), _checked_formula(
            f'=IF({status}="","",{expected_display})',
            f"Spot expected display field {f}"), body_font)
        put((r, 13), _checked_formula(
            f'=IFERROR(IF({status}="","",MID({comparison_mask},'
            f'{state_offset},1)),"")',
            f"Spot Comparison state field {f}"), body_font)
        # One-sided rows: the verdict carries the independently derived status
        # itself (tinted via CF) so it remains loud in every field row.
        put((r, 5), f'=IF({status}="","",IF($K{r}="U",'
                    f'{independent_status},IF($K{r}="D","DIFFERENT","match")))',
            body_font, None, center)
        put((r, 6), f'=IFERROR(IF(ISBLANK({comparison_display}),"",'
                    f'{comparison_display}),"")', body_font)
        put((r, 7), f'=IF({status}="","",IF(AND(EXACT($K{r},$M{r}),'
                    f'EXACT($L{r},$F{r})),"OK","CHECK"))',
            body_font, None, center)

    put((F_LAST + 2, 2),
        "• On rows that exist in only one system the verdict column shows "
        f"'{lay.only_a}' / '{lay.only_b}' on every field; Agree? then verifies the "
        "displayed value against that system's data sheet.", note_font)
    put((F_LAST + 3, 2),
        "• The blue row numbers jump to the source row on the data sheets "
        "and select the whole row so it stands out — it un-highlights when "
        "you click elsewhere. Each data-sheet row links back to its "
        "Comparison row. Values are shown exactly as stored (before TRIM).",
        note_font)
    put((F_LAST + 4, 2),
        "• Row integrity: OK means Comparison's stored row links and status "
        "exactly equal an independent MATCH of this row's hidden key token "
        "into both data sheets — the row pairing itself is verified, not "
        "assumed.", note_font)

    # Emit the sparse grid (append-only streaming sheet).
    n_rows = max(r for r, _c in grid)
    n_cols = max(c for _r, c in grid)
    for r in range(1, n_rows + 1):
        cells = []
        for c in range(1, n_cols + 1):
            if (r, c) in grid:
                value, font, fill, align, fmt = grid[(r, c)]
                cell = _styled(ws, value, font, fill, align)
                if fmt:
                    cell.number_format = fmt
                cells.append(cell)
            else:
                cells.append(None)
        ws.append(cells)
    return ws


def _write_only_sheet(wb, side, other, tab_color, keys, lay, events, vals=None,
                      helper_tokens=None):
    """'Only in <side>': every one-sided union row, in union order, with the
    full field data pulled LIVE from that system's data sheet (same
    MATCH-on-helper-key + INDEX pattern as the Comparison sheet, so the tab
    recalculates with edits). Consolidated mode adds a "Missing from <other>"
    column — "entire route" (tinted: the other system lacks the whole route)
    vs "this <noun> only" — so wholly-missing routes can't be overlooked."""
    sc = lay.sc
    name = f"Only in {side}"
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    body_font = Font(name="Arial", size=10)

    id_headers = ((["Route"] if lay.has_route else [])
                  + [sc.header[lay.key_field], "#", f"{side} Row"]
                  + ([f"Missing from {other}"] if lay.has_route else []))
    n_id = len(id_headers)
    # fields follow the id columns in display order; a header index maps to its
    # position among the displayed fields (key_field == 0 → n_id + field_idx,
    # the original mapping).
    fcol = lambda fi: get_column_letter(n_id + 1 + lay.field_pos(fi))
    first_field = get_column_letter(n_id + 1)
    last_field = get_column_letter(n_id + lay.n_fields)
    last = len(keys) + 1
    c = [get_column_letter(i + 1) for i in range(n_id)]
    if lay.has_route:
        c_route, c_loc, c_occ, c_row, c_why = c
    else:
        c_route = c_why = None
        c_loc, c_occ, c_row = c

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = f"{first_field}2"
    ws.auto_filter.ref = f"A1:{last_field}{last}"
    if lay.has_route:
        ws.column_dimensions[c_route].width = 8
        ws.column_dimensions[c_why].width = 18
    ws.column_dimensions[c_loc].width = 12
    ws.column_dimensions[c_occ].width = 4
    ws.column_dimensions[c_row].width = 9
    _apply_field_widths(ws, sc.cmp_widths, fcol, lay)
    if lay.has_route and keys:
        # Whole-route gaps stand out; single-location gaps stay plain. Colors
        # mirror the Comparison sheet's one-sided row tints. (`keys` can be
        # EMPTY when the two sides match perfectly — common between
        # environments — and an empty tab's CF range "A2:..1" is invalid.)
        tint = "FFE699" if side == sc.side_a else "BDD7EE"
        ws.conditional_formatting.add(f"A2:{last_field}{last}", FormulaRule(
            formula=[f'${c_why}2="entire route"'], fill=PatternFill(bgColor=tint)))

    link_font = _link_font()
    link_col = 3 if lay.has_route else 2              # the "<side> Row" position
    if vals is not None:
        own_rows = vals["row_t"] if side == sc.side_a else vals["row_n"]
        own_by = vals["by_t"] if side == sc.side_a else vals["by_n"]
        other_routes = vals["routes_n"] if side == sc.side_a else vals["routes_t"]
    ws.append(_header_row(ws, id_headers
                          + [sc.header[i] for i in lay.field_indices],
                          lay.sc.header_comment))
    if not isinstance(helper_tokens, dict) or any(
            key not in helper_tokens for key in keys):
        raise ValueError("Only-in helper-token map does not cover every row")
    for i, (route, loc, occ) in enumerate(keys):
        r = i + 2
        visible_loc = _visible_key(loc)
        if vals is None:
            key = _formula_text(helper_tokens[(route, loc, occ)])
            row = ([route] if lay.has_route else []) + [
                visible_loc, occ,
                _row_link(side, key, lay),
            ]
            if lay.has_route:
                rc = lay.route_data_col      # Route on the data sheets (B)
                row.append(f'=IF(COUNTIF({_sref(other)}!${rc}:${rc},$A{r})=0,'
                           f'"entire route","this {sc.id_noun} only")')
            rr = f"${c_row}{r}"
            row += [f'=IF({rr}="","",{_trim_ref(side, lay.data_col(f), rr)})'
                    for f in lay.field_indices]
        else:
            k = (route, loc, occ)
            own = own_by[k]
            row = ([route] if lay.has_route else []) + [
                visible_loc, occ,
                _row_link_value(side, own_rows[k], lay),
            ]
            if lay.has_route:
                row.append("entire route" if route not in other_routes
                           else f"this {sc.id_noun} only")
            row += [(_xl_trim(own[f + vals["off"]]) or None)
                    for f in lay.field_indices]
        # Same injection guard as the Comparison sheet: in the formulas flavor
        # only the key/route id cells are literals; in the values flavor every
        # non-link cell is. The "<side> Row" HYPERLINK must stay a formula.
        if vals is None:
            guard_set = {0} if lay.has_route else set()
            guard_set.add(1 if lay.has_route else 0)         # the key (loc) cell
        else:
            guard_set = set(range(len(row))) - {link_col}
        ws.append([_styled(ws, v, link_font if j == link_col else body_font,
                           guard=(j in guard_set))
                   for j, v in enumerate(row)])
        if (i + 1) % _PROGRESS_EVERY == 0:
            events.on_log(f"  {name} sheet: {i + 1:,} of {len(keys):,} rows…")
            if events.is_cancelled():
                return None
    return ws


def route_coverage(keys_t, keys_n, is_cancelled=None):
    """Route lists, ordered as the union orders them: side A's route order,
    then B-only routes in B order. Returns (all_routes, both, t_only,
    n_only) — each a list of route ids."""
    _raise_if_cancelled(is_cancelled)
    rt, rn = [], []
    seen_t, seen_n = set(), set()
    for key_no, k in enumerate(keys_t):
        if key_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if k[0] not in seen_t:
            seen_t.add(k[0])
            rt.append(k[0])
    for key_no, k in enumerate(keys_n):
        if key_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if k[0] not in seen_n:
            seen_n.add(k[0])
            rn.append(k[0])
    set_t, set_n = seen_t, seen_n
    all_routes = []
    for route_no, route in enumerate(rt):
        if route_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        all_routes.append(route)
    for route_no, route in enumerate(rn):
        if route_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if route not in set_t:
            all_routes.append(route)
    both, t_only, n_only = [], [], []
    for route_no, route in enumerate(all_routes):
        if route_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if route in set_t and route in set_n:
            both.append(route)
    for route_no, route in enumerate(rt):
        if route_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if route not in set_n:
            t_only.append(route)
    for route_no, route in enumerate(rn):
        if route_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        if route not in set_t:
            n_only.append(route)
    _raise_if_cancelled(is_cancelled)
    return all_routes, both, t_only, n_only


def _post_count_derivations(keys_t, keys_n, union, helper_tokens,
                            rows_t, rows_n, include_value_maps,
                            is_cancelled=None):
    """Build every pre-serialization lookup with cooperative cancellation.

    These structures used to be large comprehensions after counting, creating
    a final blind region where cancellation could arrive after exact truth was
    computed but before workbook serialization began. The returned values are
    identical; only the optional polling seam is new.
    """
    _raise_if_cancelled(is_cancelled)

    def side_maps(keys, rows):
        key_set = set()
        routes = set()
        helpers = []
        by_key = {} if include_value_maps else None
        row_by_key = {} if include_value_maps else None
        for key_no, key in enumerate(keys):
            if key_no % _PROGRESS_EVERY == 0:
                _raise_if_cancelled(is_cancelled)
            key_set.add(key)
            routes.add(key[0])
            helpers.append(helper_tokens[key])
            if include_value_maps:
                by_key[key] = rows[key_no]
                row_by_key[key] = key_no + 2
        _raise_if_cancelled(is_cancelled)
        return key_set, routes, helpers, by_key, row_by_key

    set_t, routes_t, hk_t, by_t, row_t = side_maps(keys_t, rows_t)
    set_n, routes_n, hk_n, by_n, row_n = side_maps(keys_n, rows_n)

    only_t, only_n, union_row = [], [], {}
    for key_no, key in enumerate(union):
        if key_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        union_row[key] = key_no + 2
        if key not in set_n:
            only_t.append(key)
        if key not in set_t:
            only_n.append(key)

    cmp_rows_t = []
    for key_no, key in enumerate(keys_t):
        if key_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        cmp_rows_t.append(union_row[key])
    cmp_rows_n = []
    for key_no, key in enumerate(keys_n):
        if key_no % _PROGRESS_EVERY == 0:
            _raise_if_cancelled(is_cancelled)
        cmp_rows_n.append(union_row[key])
    _raise_if_cancelled(is_cancelled)
    return {
        "only_t": only_t,
        "only_n": only_n,
        "cmp_rows_t": cmp_rows_t,
        "cmp_rows_n": cmp_rows_n,
        "hk_t": hk_t,
        "hk_n": hk_n,
        "routes_t": routes_t,
        "routes_n": routes_n,
        "by_t": by_t,
        "by_n": by_n,
        "row_t": row_t,
        "row_n": row_n,
    }


def _write_routes(wb, all_routes, lay, vals=None):
    """Consolidated mode only: one row per route with LIVE coverage stats —
    which system has it, how many rows each side carries, and how much of it
    differs. The route ids are the row universe (literals, like the
    Comparison sheet's keys); every count is a formula."""
    sc = lay.sc
    Noun = sc.id_noun_plural.capitalize()
    ws = wb.create_sheet("Routes")
    ws.sheet_properties.tabColor = _TAB_COLORS["routes"]
    body_font = Font(name="Arial", size=10)
    last = len(all_routes) + 1
    st, df = lay.c_status, lay.c_diffs

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:H{last}"
    for col, w in (("A", 8), ("B", 12), ("C", 12), ("D", 12),
                   ("E", 16), ("F", 16), ("G", 18), ("H", 14)):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30
    # Status colors match the Comparison sheet; red counts where differences exist.
    ws.conditional_formatting.add(f"A2:H{last}", FormulaRule(
        formula=[f'$B2="{lay.only_a}"'], fill=PatternFill(bgColor="FFE699")))
    ws.conditional_formatting.add(f"A2:H{last}", FormulaRule(
        formula=[f'$B2="{lay.only_b}"'], fill=PatternFill(bgColor="BDD7EE")))
    ws.conditional_formatting.add(f"G2:H{last}", CellIsRule(
        operator="greaterThan", formula=["0"],
        font=Font(color="C00000", bold=True)))

    ws.append(_header_row(ws, [
        "Route", "Status", f"{sc.side_a} rows", f"{sc.side_b} rows",
        f"{Noun} compared", f"Matched {sc.id_noun_plural}",
        f"{Noun} w/ differences", "Differing cells"]))
    rc = lay.route_data_col                  # Route on the data sheets (B)
    sa, sb = _sref(sc.side_a), _sref(sc.side_b)
    for i, route in enumerate(all_routes):
        r = i + 2
        if vals is None:
            cells = (
                route,
                f'=IF(AND(C{r}>0,D{r}>0),"Both",IF(C{r}>0,"{lay.only_a}","{lay.only_b}"))',
                f'=COUNTIF({sa}!${rc}:${rc},$A{r})',
                f'=COUNTIF({sb}!${rc}:${rc},$A{r})',
                f'=COUNTIF(Comparison!$A:$A,$A{r})',
                f'=COUNTIFS(Comparison!$A:$A,$A{r},Comparison!${st}:${st},"Both")',
                f'=COUNTIFS(Comparison!$A:$A,$A{r},Comparison!${st}:${st},"Both",'
                f'Comparison!${df}:${df},">0")',
                f'=SUMIF(Comparison!$A:$A,$A{r},Comparison!${df}:${df})',
            )
        else:
            rs = vals["counts"]["route"][route]
            cells = (
                route,
                ("Both" if rs["t_rows"] and rs["n_rows"]
                 else lay.only_a if rs["t_rows"] else lay.only_b),
                rs["t_rows"], rs["n_rows"], rs["locs"], rs["matched"],
                rs["withdiffs"], rs["cells"],
            )
        # Guard only the route-id cell (cell 0): a route like "=X" would become
        # a live formula. The other cells are our own =formulas (formulas flavor)
        # or safe literals (values flavor) and must NOT be forced to text.
        ws.append([_styled(ws, v, body_font, guard=(j == 0))
                   for j, v in enumerate(cells)])
    return ws


def _build_snapshot_freshness_expr(lay, source_row_counts):
    """Excel predicate proving both source sheets still equal their snapshots."""
    if (not isinstance(source_row_counts, tuple) or len(source_row_counts) != 2
            or any(type(value) is not int or value < 0
                   for value in source_row_counts)):
        raise ValueError("source_row_counts must be two non-negative integers")
    terms = []
    for side, expected_count in zip(
            (lay.sc.side_a, lay.sc.side_b), source_row_counts):
        data = _sref(side)
        snapshot = _sref(lay.snapshot_sheet(side))
        last = expected_count + 1
        terms.append(
            f"COUNTA({data}!${lay.back_col}:${lay.back_col})-1={expected_count}")
        terms.append(f"COUNTA({snapshot}!$A:$A)-1={expected_count}")
        for column in lay.build_fresh_cols:
            cells = f"{data}!${column}$2:${column}${last}"
            terms.append(f'COUNTIF({cells},"STALE")=0')
            terms.append(
                f'COUNTIF({cells},"OK")={expected_count}')
            terms.append(f'{data}!${column}${expected_count + 2}="END"')
    return f"AND({','.join(terms)})"


def _write_summary(wb, name_a, name_b, n_union, lay, vals=None, warnings=(),
                   pairing_diagnostics=(), source_row_counts=()):
    """`vals` None = live-formula stats; else the values model and every
    stat is its literal number. The SELF-CHECK rows stay LIVE in both
    flavors — in the values workbook they recount the literal sheets, so
    they still prove the written numbers are internally consistent.

    `warnings` (input files that were unreadable/skipped on one or both sides)
    makes the report honest about incompleteness: a clean "everything matches"
    is downgraded to a loud "could not compare everything" banner and the skips
    are listed in the notes. ``pairing_diagnostics`` independently marks an
    above-cap positional fallback: its observed counts remain visible, but the
    workbook cannot certify a match or exact row identity."""
    sc = lay.sc
    a, b = sc.side_a, sc.side_b
    noun, nouns = sc.id_noun, sc.id_noun_plural
    Nouns = nouns.capitalize()
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = _TAB_COLORS["summary"]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20

    title_font = Font(name="Arial", size=14, bold=True, color=_DARK)
    note_font = Font(name="Arial", size=10, color="595959")
    banner_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    banner_fill = PatternFill("solid", start_color=_DARK)
    body_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    center = Alignment(horizontal="center")
    last = n_union + 1                              # Comparison data end row
    loc_col = lay.data_col(lay.key_field)  # key column on the data sheets
    st, df = lay.c_status, lay.c_diffs
    sa, sb = _sref(a), _sref(b)

    # Streaming sheets are append-only: build a sparse (row, col) grid first,
    # emit it row by row at the end.
    grid = {}
    row = [1]                                       # 1-slot mutable cursor

    def put(col, value, font=body_font, fill=None, align=None):
        grid[(row[0], col)] = (value, font, fill, align)

    def line(*cells, advance=1):
        for col, value, *style in cells:
            put(col, value, *style)
        row[0] += advance

    def banner(text):
        line((2, text, banner_font, banner_fill))

    def stat(label, formula, value=None):
        line((2, label),
             (3, formula if vals is None else value, bold_font, None, center))

    c = None if vals is None else vals["counts"]
    scope = (sc.scope_consolidated if lay.has_route else sc.scope_flat) \
        + ("" if vals is None else " — VALUES copy")
    manual_banner = lay.has_route and vals is None
    # THE VERDICT — the one-line answer most comparisons exist for ("did
    # anything change between the two sides?"). Live in the formulas flavor
    # (recomputes with edits / after F9; blank until then in manual mode), a
    # literal in the values flavor. Green/red via CF keyed on the ✓/✗ glyph.
    verdict_row = 4 if manual_banner else 3
    verdict_font = Font(name="Arial", size=12, bold=True)
    ws.conditional_formatting.add(f"B{verdict_row}", FormulaRule(
        formula=[f'LEFT($B${verdict_row},1)="✓"'],
        fill=PatternFill(bgColor="C6EFCE"), font=Font(color="2E7D32", bold=True)))
    ws.conditional_formatting.add(f"B{verdict_row}", FormulaRule(
        formula=[f'LEFT($B${verdict_row},1)="✗"'],
        fill=PatternFill(bgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))
    match_text = (f"✓ EVERYTHING MATCHES — all {n_union:,} {nouns} are "
                  f"identical in both {sc.sides_noun}.")
    pairing_capped = bool(pairing_diagnostics)
    if warnings and pairing_capped:
        match_text = (
            "✗ COULD NOT CERTIFY A COMPLETE COMPARISON — some inputs were "
            "unreadable and at least one duplicate group exceeded the exact-"
            "pairing cap. See the notes; observed counts are non-certifying.")
    elif pairing_capped:
        match_text = (
            "✗ PAIRING LIMIT REACHED — all displayed counts use deterministic "
            "positional fallback for at least one duplicate group. Row identity "
            "is unresolved; this workbook cannot certify a match. See the notes.")
    elif warnings:
        # Some inputs couldn't be read — a clean "match" can't be certified.
        # Leads with ✗ so the existing red conditional-format applies (no new
        # CF rule, so the no-warnings path stays byte-identical).
        match_text = (f"✗ COULD NOT COMPARE EVERYTHING — {len(warnings)} input "
                      f"file(s) were unreadable and skipped; of the {n_union:,} "
                      f"{nouns} that WERE compared, all match. See the notes.")
    if vals is None:
        diff_cells_ref = f"SUM(Comparison!{df}2:{df}{last})"
        one_sided_ref = (f'COUNTIF(Comparison!{st}:{st},"{lay.only_a}")'
                         f'+COUNTIF(Comparison!{st}:{st},"{lay.only_b}")')
        if pairing_capped:
            prefix = ("✗ PARTIAL / PAIRING LIMIT + INPUT GAPS — "
                      if warnings else "✗ PARTIAL / PAIRING LIMIT — ")
            verdict_cell = (
                f'="{prefix}positional fallback observed "&'
                f'TEXT({diff_cells_ref},"#,##0")&" differing cell(s), "&'
                f'TEXT({one_sided_ref},"#,##0")&" one-sided row(s). "&'
                '"These are diagnostic counts, not certified differences; "&'
                '"re-scope and regenerate."')
        else:
            verdict_cell = (
                f'=IF(AND({diff_cells_ref}=0,{one_sided_ref}=0),"{match_text}",'
                f'"✗ DIFFERENCES FOUND — "&TEXT({diff_cells_ref},"#,##0")&'
                f'" differing cell(s), "&TEXT({one_sided_ref},"#,##0")&'
                f'" one-sided row(s) — details below.")')
    else:
        one_sided = c["t_only"] + c["n_only"]
        if pairing_capped:
            prefix = ("✗ PARTIAL / PAIRING LIMIT + INPUT GAPS — "
                      if warnings else "✗ PARTIAL / PAIRING LIMIT — ")
            verdict_cell = (
                f"{prefix}positional fallback observed {c['diff_cells']:,} "
                f"differing cell(s), {one_sided:,} one-sided row(s). These are "
                "diagnostic counts, not certified differences; re-scope and "
                "regenerate.")
        else:
            verdict_cell = (match_text if c["diff_cells"] == 0 and one_sided == 0
                            else f"✗ DIFFERENCES FOUND — {c['diff_cells']:,} "
                                 f"differing cell(s), {one_sided:,} one-sided "
                                 "row(s) — details below.")
    freshness = _build_snapshot_freshness_expr(lay, source_row_counts)
    stale_text = (
        "✗ REGENERATE REQUIRED — a source/helper row changed after this "
        "workbook was built. Displayed counts are observations under stale "
        "build-time identity/pairing and are not certified.")
    if vals is None:
        if not isinstance(verdict_cell, str) or not verdict_cell.startswith("="):
            raise ValueError("live Summary verdict must be an Excel formula")
        verdict_cell = (
            f'=IF({freshness},{verdict_cell[1:]},{_formula_text(stale_text)})')
    else:
        verdict_cell = (
            f'=IF({freshness},{_formula_text(verdict_cell)},'
            f'{_formula_text(stale_text)})')
    row[0] = 2
    line((2, f"{a} vs {b} — {sc.report_name} — Discrepancy Report ({scope})", title_font))
    if manual_banner:
        # The big formulas workbook ships UNCALCULATED (manual mode): every
        # cell shows blank/0 until F9. Without a loud banner that reads as
        # broken data. (The values copy calculates nothing — no banner.)
        line((2, "▶ PRESS F9 TO CALCULATE — this workbook opens uncalculated "
                 "(blank/0 cells). The first F9 takes a few minutes; let it "
                 "finish, then save.",
              Font(name="Arial", size=11, bold=True, color="C00000")))
    assert row[0] == verdict_row            # the CF above targets this cell
    line((2, verdict_cell, verdict_font))
    line((2, "Cell-by-cell comparison keyed on "
             + (f"Route + {sc.header[lay.key_field]}" if lay.has_route
                else sc.header[lay.key_field])
             + " (+ occurrence for duplicates). "
             + (f"Core Comparison/Summary/Spot formulas recalculate observed "
                f"values after edits on the {a} / {b} sheets, but row identity, "
                "duplicate assignment, and familiar views are build-time state. "
                "ANY source/helper edit makes the Summary say REGENERATE "
                "REQUIRED; only a newly generated workbook is certifying."
                if vals is None else
                "This copy holds plain VALUES — it opens instantly and "
                "nothing needs calculating, but edits do NOT recalculate "
                "(the live-formulas copy does that). The Spot Check sheet "
                "and the SELF-CHECK rows below stay live."), note_font))
    line((2, f"{a}: {name_a}      {b}: {name_b}      "
             f"created {date.today().isoformat()}", note_font), advance=2)

    banner("ROW COUNTS")
    stat(f"{a} data rows", f"=COUNTA({sa}!{lay.back_col}:{lay.back_col})-1",
         None if vals is None else vals["n_t"])
    stat(f"{b} data rows", f"=COUNTA({sb}!{lay.back_col}:{lay.back_col})-1",
         None if vals is None else vals["n_n"])
    stat(f"Union of {nouns} compared",
         f"=COUNT(Comparison!{lay.c_occ}:{lay.c_occ})", n_union)
    banner("MATCH STATUS")
    stat(f"{Nouns} in both {sc.sides_noun}", f'=COUNTIF(Comparison!{st}:{st},"Both")',
         c and c["both"])
    stat(f"In {a} only (missing from {b}) — listed on the 'Only in {a}' sheet",
         f'=COUNTIF(Comparison!{st}:{st},"{lay.only_a}")', c and c["t_only"])
    stat(f"In {b} only (missing from {a}) — listed on the 'Only in {b}' sheet",
         f'=COUNTIF(Comparison!{st}:{st},"{lay.only_b}")', c and c["n_only"])
    if lay.has_route:
        banner("ROUTE COVERAGE (see the Routes sheet for the per-route breakdown)")
        stat(f"Routes covered by both {sc.sides_noun}", '=COUNTIF(Routes!B:B,"Both")',
             None if vals is None else vals["r_both"])
        stat(f"Routes only in {a} (missing from {b})",
             f'=COUNTIF(Routes!B:B,"{lay.only_a}")',
             None if vals is None else vals["r_t_only"])
        stat(f"Routes only in {b} (missing from {a})",
             f'=COUNTIF(Routes!B:B,"{lay.only_b}")',
             None if vals is None else vals["r_n_only"])
    banner("FIELD-LEVEL DISCREPANCIES (matched rows)")
    stat("Matched rows with ≥ 1 field difference",
         f'=COUNTIFS(Comparison!{st}2:{st}{last},"Both",'
         f'Comparison!{df}2:{df}{last},">0")', c and c["diff_rows"])
    stat("Matched rows fully identical",
         f'=COUNTIFS(Comparison!{st}2:{st}{last},"Both",'
         f'Comparison!{df}2:{df}{last},0)', c and c["identical"])
    stat("Total differing cells", f"=SUM(Comparison!{df}2:{df}{last})",
         c and c["diff_cells"])
    row[0] += 1

    banner("DIFFERENCES BY FIELD")
    line((2, "Field", bold_font), (3, "Comparison col", bold_font),
         (4, "# of cells differing", bold_font))
    f_start = row[0]
    for f in lay.field_indices:
        col = lay.field_col(f)
        chunk, offset = lay.state_location(f)
        state_range = (f'Comparison!${chunk["col"]}$2:'
                       f'${chunk["col"]}${last}')
        line((2, sc.header[f]), (3, col),
             (4, f'=SUMPRODUCT(--(MID({state_range},{offset},1)="D"))'
              if vals is None else c["field_diffs"][f],
              bold_font, None, center))
    f_end = row[0] - 1
    row[0] += 1

    # Live cross-checks: each headline number recomputed a second, independent
    # way. Any row reading CHECK means a formula no longer points where it
    # should (the classic cause: rows inserted/deleted on a data sheet).
    banner("SELF-CHECK (every row should read OK after calculation)")
    only_occ = "C" if lay.has_route else "B"          # occurrence (#) col on Only-in tabs
    only_a_ref = _sref(f"Only in {a}")
    only_b_ref = _sref(f"Only in {b}")

    def check(label, cond):
        line((2, label), (3, f'=IF({cond},"OK","CHECK")', bold_font, None, center))

    union_count = f"COUNT(Comparison!{lay.c_occ}:{lay.c_occ})"
    check(f"Every Comparison row has a status (Both + {lay.only_a} + {lay.only_b})",
          f'COUNTIF(Comparison!{st}:{st},"Both")'
          f'+COUNTIF(Comparison!{st}:{st},"{lay.only_a}")'
          f'+COUNTIF(Comparison!{st}:{st},"{lay.only_b}")={union_count}')
    check(f"Every row with {a} data found its {a} sheet row",
          f'COUNT(Comparison!{lay.c_trow}:{lay.c_trow})='
          f'COUNTIF(Comparison!{st}:{st},"Both")'
          f'+COUNTIF(Comparison!{st}:{st},"{lay.only_a}")')
    check(f"Every row with {b} data found its {b} sheet row",
          f'COUNT(Comparison!{lay.c_nrow}:{lay.c_nrow})='
          f'COUNTIF(Comparison!{st}:{st},"Both")'
          f'+COUNTIF(Comparison!{st}:{st},"{lay.only_b}")')
    check(f"'Only in {a}' sheet rows = {a}-only rows in the Comparison",
          f"COUNT({only_a_ref}!{only_occ}:{only_occ})="
          f'COUNTIF(Comparison!{st}:{st},"{lay.only_a}")')
    check(f"'Only in {b}' sheet rows = {b}-only rows in the Comparison",
          f"COUNT({only_b_ref}!{only_occ}:{only_occ})="
          f'COUNTIF(Comparison!{st}:{st},"{lay.only_b}")')
    check("Per-field difference counts add up to the total differing cells",
          f'SUM(D{f_start}:D{f_end})=SUM(Comparison!{df}2:{df}{last})')
    check("Build-time source identity and duplicate pairing snapshot is current",
          freshness)
    if sc.report_view_diff_check:
        cfg = sc.report_view_diff_check
        valid_cfg = (
            isinstance(cfg, tuple) and len(cfg) == 3
            and isinstance(cfg[0], str) and bool(cfg[0])
            and isinstance(cfg[1], str)
            and re.fullmatch(r"[A-Z]{1,3}", cfg[1]) is not None
            and type(cfg[2]) is int and cfg[2] > 0
            and sc.extra_sheet_writer is not None
        )
        if not valid_cfg:
            raise ValueError(
                "report_view_diff_check must be (sheet name, uppercase Diffs "
                "column, positive repeat count) and requires extra_sheet_writer")
        rv_sheet, rv_col, repeat = cfg
        check("Report View Diffs agree with the Comparison",
              f"SUM({_sref(rv_sheet)}!{rv_col}:{rv_col})="
              f"{repeat}*SUM(Comparison!{df}2:{df}{last})")
    if lay.has_route:
        check(f"Routes sheet {a} row counts add up to the {a} sheet",
              f"SUM(Routes!C:C)=COUNTA({sa}!{lay.back_col}:{lay.back_col})-1")
        check(f"Routes sheet {b} row counts add up to the {b} sheet",
              f"SUM(Routes!D:D)=COUNTA({sb}!{lay.back_col}:{lay.back_col})-1")
        check(f"Routes sheet '{Nouns} compared' adds up to the Comparison",
              f"SUM(Routes!E:E)={union_count}")
    row[0] += 1

    banner("HOW TO READ / NOTES")
    notes = [
        "• Comparison sheet: matching values are shown in plain text; a red "
        f"cell shows  {a} value{_DIFF_MARK}{b} value  where the two {sc.sides_noun} "
        f"disagree for that {sc.header[lay.key_field]} and field.",
        '• "(blank)" means the cell is empty in that system. Filter the Diffs '
        "column (>0) to isolate rows needing review.",
        f"• Yellow rows exist only in {a}; blue rows exist only in {b}"
        f"{sc.one_sided_note_extra}. Their "
        "field cells show that system's own values.",
        f"• The 'Only in {a}' and 'Only in {b}' sheets repeat every "
        "one-sided row in one place — including the rows of routes the other "
        "system doesn't carry at all"
        + (" (flagged 'entire route' and tinted; filter the 'Missing from …' "
           "column to separate whole-route gaps from single locations)"
           if lay.has_route else "")
        + ". The Comparison sheet still contains the same rows in document "
        "order.",
        "• Rows pair on " + ("Route plus " if lay.has_route else "")
        + f"{sc.header[lay.key_field]} plus occurrence number. When a "
        f"{sc.pair_noun or noun} is listed more than once, the matching "
        "instances are paired by which are MOST ALIKE (fewest differing "
        "fields), not by the order they appear — so a repeat that matches the "
        "other side's second listing isn't flagged as a difference.",
        f"• Leading/trailing spaces are ignored (TRIM){sc.trim_note_extra}.",
        f'• Lookups use the "Key (helper)" column ({lay.key_col}) on each '
        "data sheet. It contains a versioned opaque build token, not flattened "
        "Route/key text, so punctuation inside identity components cannot collide.",
        "• Very-hidden E2 snapshot sheets retain the exact build-time source and "
        "helper cells. The Summary freshness check turns non-certifying after any "
        "source edit, row insertion/deletion/reorder, or helper change.",
    ]
    if warnings:
        shown = warnings[:20]
        notes.append(
            f"• ⚠ INCOMPLETE COMPARISON — {len(warnings)} input file(s) could not "
            "be read and were left OUT of this comparison, so anything in them "
            "is neither matched nor flagged. Re-export/repair them and re-run "
            "before trusting a clean result:")
        for w in shown:
            notes.append(f"     – {w}")
        if len(warnings) > len(shown):
            notes.append(f"     – …and {len(warnings) - len(shown)} more "
                         "(see the run log).")
    if pairing_capped:
        notes.append(
            f"• ⚠ PAIRING INCOMPLETE — {len(pairing_diagnostics)} duplicate "
            "key group(s) exceeded the 100,000-cell exact-assignment cap. "
            "Those groups use deterministic positional pairing only; all "
            "displayed counts remain useful diagnostics but are not certified "
            "minimum-cost results. Re-scope the comparison and re-run before "
            "treating it as a match.")
        for diagnostic in pairing_diagnostics[:20]:
            key = " / ".join(diagnostic.key_components)
            notes.append(
                f"     – {key}: {diagnostic.side_a_size} × "
                f"{diagnostic.side_b_size} rows; positional fallback cost "
                f"{diagnostic.fallback_cost}")
        if len(pairing_diagnostics) > 20:
            notes.append(
                f"     – …and {len(pairing_diagnostics) - 20} more capped "
                "groups (see the typed outcome sidecar).")
    if sc.medwid_fields:
        notes.append(
            "• Med Wid is compared after normalizing zero-padding in the numeric "
            f"part ({a} 0Z = {b} 00Z, 6V = 06V, etc.), since the two systems "
            "format this code differently. All other fields compare exactly.")
    notes.append(
        f"• Doubting a value? The blue row numbers in the '{a} Row' / "
        f"'{b} Row' columns are clickable — they jump to the data sheet and "
        "SELECT that whole row (it stays highlighted until you click "
        "elsewhere), and each data-sheet row links back to its Comparison "
        f"row the same way. The Spot Check sheet audits any single {noun} "
        "end to end: raw values from both systems and an independently "
        "recomputed verdict for every field.")
    if vals is not None:
        notes.append(
            "• This is the VALUES copy: every number and comparison cell is "
            "a computed result, not a formula (only the Spot Check sheet and "
            "the SELF-CHECK rows stay live). If the data changes, re-create "
            "the comparison — or use the live-formulas copy, which "
            "recalculates.")
    notes.append(
        "• SELF-CHECK recomputes the headline numbers a second, independent "
        "way; a CHECK there means the sheets no longer agree (e.g. rows were "
        "inserted or deleted on a data sheet) — re-create the report rather "
        "than trust the numbers.")
    if lay.has_route:
        notes.append(
            "• The Routes sheet lists every route either system carries — "
            "which side covers it, row counts, and how much of it differs.")
        if vals is None:
            notes.append(
                "• CALCULATION IS SET TO MANUAL (large workbook): cells show "
                "blank/0 until you press F9. The first F9 takes a few minutes — "
                "let it finish, then save to keep the results; edits afterwards "
                "only recalculate when you press F9 again. (Excel keeps the "
                "manual setting for other workbooks opened in the same session — "
                "Formulas → Calculation Options switches it back.)")
    for note in notes:
        line((2, note, note_font))

    # Emit the grid (append-only streaming sheet).
    n_rows = max(r for r, _c in grid)
    n_cols = max(c for _r, c in grid)
    for r in range(1, n_rows + 1):
        cells = []
        for c in range(1, n_cols + 1):
            if (r, c) in grid:
                value, font, fill, align = grid[(r, c)]
                cells.append(_styled(ws, value, font, fill, align))
            else:
                cells.append(None)
        ws.append(cells)


def _write_provenance_sheet(wb, provenance):
    """Append the small human-facing "Provenance" sheet (CMP-AUD-076): what the
    comparison actually consumed — the recipe, each input's role/kind, its FULL
    canonical selection, and its identity facts (content digest for a file; the
    discovered member count for a folder) — captured BEFORE the inputs were
    read. Write-only-safe (create_sheet + append only). The machine-readable
    record incl. the committed generation binding is the `.provenance.json`
    sidecar beside the workbook; this sheet is its concise display."""
    ws = wb.create_sheet("Provenance")
    ws.sheet_properties.tabColor = "808080"
    for col, w in (("A", 12), ("B", 110)):
        ws.column_dimensions[col].width = w
    recipe = provenance.get("recipe") or {}
    ws.append(["Comparison Provenance"])
    ws.append(["What this workbook compared — captured before the inputs "
               "were read."])
    ws.append([])
    ws.append(["Report", recipe.get("report", "")])
    ws.append(["Run", recipe.get("banner", "")])
    for rec in provenance.get("inputs") or ():
        ws.append([])
        ws.append([str(rec.get("role", "")), str(rec.get("selection", ""))])
        if rec.get("kind") == "folder":
            ws.append(["", f"{rec.get('member_count', 0)} discovered source "
                           "file(s); the exact member census is in the "
                           "provenance sidecar"])
        else:
            ws.append(["", f"sha256 {rec.get('sha256', '')}"])
        if rec.get("producer_completion") is not None:
            ws.append(["", f"producer completion: {rec['producer_completion']}"])
    ws.append([])
    ws.append(["Note", "The machine-readable record (including the committed "
                       "generation binding) is the .provenance.json sidecar "
                       "beside this workbook."])


def run_compare(sc, rows_t, rows_n, has_route, out_path, *, events=None,
                confirm_overwrite=None, mode="formulas",
                name_a="", name_b="", warnings=(), commit_guard=None,
                input_completion=None, skipped_inputs=None, failed_inputs=0,
                failures=(), coverage_diagnostics=(), provenance=None):
    """Build the comparison workbook(s) from two loaded row sets. Returns a
    ConsolidateResult (same contract as the consolidators, so the GUI/console
    drive it identically). The CALLER owns input loading + shape validation;
    `rows_t`/`rows_n` are side A / side B in the schema's column order.

    `mode`: "formulas" (the live workbook — every cell recalculates),
    "values" (same sheets and look, but the bulk is plain computed RESULTS —
    opens instantly, no F9; links, conditional formatting, the Spot Check
    sheet and the SELF-CHECK rows are kept), or "both" (two files: the picked
    name for the formulas copy and '<name> (values).xlsx' next to it).

    `warnings`: input files the caller couldn't read (one per string). When
    non-empty the comparison is INCOMPLETE — the verdict can never be a clean
    "match" (forced to "diff"), the workbook banner says so, and the skipped
    files are listed. Empty → unchanged (the regression-locked default).

    ``input_completion`` and the explicit issue counters are the additive
    Phase-2 bridge for loaders that already own structured coverage truth.
    Their defaults preserve the legacy ``skipped_inputs == len(warnings)``
    behavior; exact counts are never derived from display prose."""
    warnings = [str(w) for w in warnings]
    structured_warnings = list(warnings)
    failure_items = tuple(str(item) for item in (failures or ()))
    structured_coverage = tuple(coverage_diagnostics or ())
    pairing_trace = ()
    pairing_quality = "unknown"
    pairing_diagnostics = ()
    input_meta_error = None
    try:
        if skipped_inputs is None:
            exact_skipped = len(warnings)
        elif isinstance(skipped_inputs, bool) or not isinstance(skipped_inputs, int):
            raise ValueError
        else:
            exact_skipped = skipped_inputs
        if isinstance(failed_inputs, bool) or not isinstance(failed_inputs, int):
            raise ValueError
        exact_failed = failed_inputs
        if exact_skipped < 0 or exact_failed < 0:
            raise ValueError
    except (TypeError, ValueError):  # silent-ok: malformed typed input becomes a fail-closed error result below
        exact_skipped = exact_failed = 0
        input_meta_error = "Comparison input issue counts must be non-negative integers."
    if input_completion not in (None, outcome.COMPLETE, outcome.PARTIAL):
        input_meta_error = "Comparison input completion must be complete, partial, or absent."
    input_incomplete = bool(
        warnings or failure_items or exact_skipped or exact_failed
        or input_completion == outcome.PARTIAL)
    # Structured failures must also be visible in the workbook/log. Keep the
    # typed failure tuple separate: display prose never owns the exact counts.
    for item in failure_items:
        if item not in warnings:
            warnings.append(item)
    if input_incomplete and not warnings:
        generated_warning = (
            "An input producer reported incomplete coverage without an "
            "itemized diagnostic.")
        warnings.append(generated_warning)
        structured_warnings.append(generated_warning)
    events = events or Events()
    a, b = sc.side_a, sc.side_b

    def _comparison_counts(raw):
        """Return metadata-only typed counts without changing workbook data."""
        if raw is None:
            return ComparisonCounts()
        # Include the stable schema index so duplicate display labels remain
        # injective instead of silently overwriting one another.
        per_field = {
            f"{field_idx}:{sc.header[field_idx]}": int(value)
            for field_idx, value in raw.get("field_diffs", {}).items()
        }
        return ComparisonCounts(
            known=True,
            paired_rows=int(raw["both"]),
            side_a_only_rows=int(raw["t_only"]),
            side_b_only_rows=int(raw["n_only"]),
            differing_rows=int(raw["diff_rows"]),
            differing_cells=int(raw["diff_cells"]),
            per_field_counts=per_field,
            asserted_cells=int(raw.get("asserted_cells", 0)),
            context_cells=int(raw.get("context_cells", 0)),
        )

    def _result(*, status="ok", message="", output_path="", summary_lines=None,
                verdict=None, completion=None, raw_counts=None,
                warning_items=(), failure_items=(), skipped_inputs=0,
                failed_inputs=0):
        """Build one normalized legacy+typed comparison result."""
        if completion is None:
            completion = (outcome.CANCELLED if status == "cancelled"
                          else outcome.FAILED if status == "error"
                          else outcome.COMPLETE)
        failures = tuple(str(item) for item in failure_items)
        if status == "error" and not failures and message:
            failures = (str(message),)
        result = ConsolidateResult(
            status=status,
            message=message,
            output_path=output_path,
            summary_lines=list(summary_lines or ()),
            verdict=verdict,
            completion=completion,
            skipped_inputs=int(skipped_inputs or 0),
            failed_inputs=int(failed_inputs or 0),
        )
        result.comparison_outcome = comparison_outcome_from_legacy(
            result,
            counts=_comparison_counts(raw_counts),
            warnings=tuple(str(item) for item in warning_items),
            failures=failures,
            pairing_trace=(pairing_trace if completion in (
                outcome.COMPLETE, outcome.PARTIAL) else ()),
            duplicate_group_count=(len(pairing_trace) if completion in (
                outcome.COMPLETE, outcome.PARTIAL) else 0),
            pairing_quality=(pairing_quality if completion in (
                outcome.COMPLETE, outcome.PARTIAL) else "unknown"),
            capped_group_diagnostics=(pairing_diagnostics if completion in (
                outcome.COMPLETE, outcome.PARTIAL) else ()),
            coverage_diagnostics=structured_coverage,
        )
        return result

    if not _DEPS_OK:
        return _result(status="error",
                       message="Required components are missing (openpyxl).")
    if input_meta_error:
        return _result(status="error", message=input_meta_error)
    confirm = confirm_overwrite or (lambda _p: True)
    out = Path(out_path)
    computed_counts = None

    def output_allowed(path):
        """Target-aware lease seam for the real serializer boundary.

        Public adapters provide a ``guard(path)`` callback. A zero-argument
        predicate cannot authorize this exact serializer target and fails
        closed; artifact_store independently binds the exclusive temp's
        ordinary-file identity before this engine receives it.
        """
        if commit_guard is None:
            return True
        try:
            return bool(commit_guard(Path(path)))
        except Exception:  # silent-ok: any guard defect is the fail-closed denial
            return False

    def ownership_error():
        return _result(
            status="error",
            message=("Refusing to write the comparison: destination ownership "
                     "changed while the workbook was being built. No output "
                     "was published; re-run after the destination is stable."),
            raw_counts=computed_counts,
            warning_items=structured_warnings,
            failure_items=failure_items,
            skipped_inputs=exact_skipped,
            failed_inputs=exact_failed)

    def close_unsaved(workbook):
        """Finalize write-only sheet generators before abandoning a guarded save."""
        for worksheet in workbook.worksheets:
            worksheet.close()
        workbook.close()

    sheets = sc.sheet_names(has_route)
    for sheet in sheets:
        if len(sheet) > 31:
            return _result(
                status="error",
                message=(f"The sheet name '{sheet}' is longer than Excel's "
                         "31-character limit — shorten the side labels."))
    # Excel sheet names must be unique case-insensitively. A side literally named
    # like a fixed sheet ("Summary"/"Comparison"/"Routes"/"Only in …") would
    # collide and openpyxl would raise mid-write; fail early with guidance.
    seen_sheets = {}
    for sheet in sheets:
        low = sheet.casefold()
        if low in seen_sheets:
            return _result(
                status="error",
                message=(f"The side name '{sheet}' collides with another sheet "
                         f"('{seen_sheets[low]}') — pick different side labels "
                         "(avoid Summary / Comparison / Routes / Spot Check / "
                         "'Only in …')."))
        seen_sheets[low] = sheet

    modes = {"formulas": ("formulas",), "values": ("values",),
             "both": ("formulas", "values")}.get(mode)
    if modes is None:
        return _result(status="error",
                       message=f"Unknown comparison mode: {mode}")
    out_paths = {m: out for m in modes}
    if len(modes) > 1:                  # the values twin sits next to the pick
        out_paths["values"] = out.with_name(f"{out.stem} (values){out.suffix}")

    if not rows_t or not rows_n:
        return _result(
            status="error",
            message="One of the inputs has no data rows — nothing to compare.",
            warning_items=structured_warnings,
            failure_items=failure_items,
            skipped_inputs=exact_skipped,
            failed_inputs=exact_failed)
    if events.is_cancelled():
        return _result(status="cancelled", message="Cancelled by user.",
                       warning_items=structured_warnings,
                       failure_items=failure_items,
                       skipped_inputs=exact_skipped,
                       failed_inputs=exact_failed)

    # openpyxl serializes NaN and infinities as blank numeric cells. That would
    # make Python/values truth disagree with the live formulas and can even turn
    # a source anomaly into apparent equality. Reject the public comparison
    # before output guards/prompts, layout, counting, or workbook writes.
    scanned_source_cells = 0
    for side, rows in ((a, rows_t), (b, rows_n)):
        for row_offset, source_row in enumerate(rows):
            if row_offset % _PROGRESS_EVERY == 0 and events.is_cancelled():
                return _result(status="cancelled", message="Cancelled by user.",
                               warning_items=structured_warnings,
                               failure_items=failure_items,
                               skipped_inputs=exact_skipped,
                               failed_inputs=exact_failed)
            row_index = row_offset + 2
            for column_index, value in enumerate(source_row, start=1):
                if (scanned_source_cells % _PROGRESS_EVERY == 0
                        and events.is_cancelled()):
                    return _result(
                        status="cancelled", message="Cancelled by user.",
                        warning_items=structured_warnings,
                        failure_items=failure_items,
                        skipped_inputs=exact_skipped,
                        failed_inputs=exact_failed)
                scanned_source_cells += 1
                if _is_nonfinite_source_numeric(value):
                    return _result(
                        status="error",
                        message=(f"{side} row {row_index}, column {column_index} "
                                 "contains a non-finite numeric value. Replace "
                                 "NaN/infinity with an explicit finite value or "
                                 "text before comparing."),
                        warning_items=structured_warnings,
                        failure_items=failure_items,
                        skipped_inputs=exact_skipped,
                        failed_inputs=exact_failed)

    # Reject obviously over-wide schemas before _Layout asks openpyxl to turn
    # beyond-XFD helper indices into column letters. The exact Comparison-mask
    # width is checked again after chunk planning; this raw bound covers visible
    # columns and the deterministic five-per-Med-Wid data helpers.
    medwid_count = sum(1 for f in sc.field_indices if sc.is_medwid(f))
    raw_data_cols = (len(sc.header) + (1 if has_route else 0) + 2
                     + len(_MEDWID_HELPER_STAGES) * medwid_count)
    raw_comparison_cols = (7 if has_route else 6) + sc.n_fields
    raw_spot_cols = 29 if medwid_count else 13
    raw_limit_err = excel_limit_error(
        1, max(raw_data_cols, raw_comparison_cols, raw_spot_cols))
    if raw_limit_err:
        return _result(status="error", message=raw_limit_err)

    for path in out_paths.values():
        if not output_allowed(path.parent) or not output_allowed(path):
            return ownership_error()

    for m in modes:
        if out_paths[m].exists() and not confirm(out_paths[m]):
            return _result(status="cancelled",
                           message="Cancelled. Existing file kept.",
                           warning_items=structured_warnings,
                           failure_items=failure_items,
                           skipped_inputs=exact_skipped,
                           failed_inputs=exact_failed)

    try:
        lay = _Layout(sc, has_route)
    except ValueError as exc:
        return _result(status="error", message=str(exc))
    # Pair duplicate keys by data similarity, not file order, so a row that
    # actually matches the other side's SECOND instance isn't reported as a
    # difference against its FIRST (can only remove phantom diffs, never add).
    try:
        keys_t = keys_for(
            rows_t, has_route, sc.key_field, sc.key_normalizer,
            events.is_cancelled)
        keys_n = keys_for(
            rows_n, has_route, sc.key_field, sc.key_normalizer,
            events.is_cancelled)
        pairing = pair_occurrences_by_similarity(
            sc, rows_t, rows_n, keys_t, keys_n, has_route, events)
        # Close the hand-off race before any partial pairing state is copied to
        # result-owned variables or used to build counts/output.
        _raise_if_cancelled(events.is_cancelled)
        keys_t, keys_n = pairing.keys_a, pairing.keys_b
        union = union_keys(keys_t, keys_n, events.is_cancelled)
        helper_tokens = _opaque_helper_tokens(
            union, events.is_cancelled)
        _raise_if_cancelled(events.is_cancelled)
    except RunCancelled:
        return _result(status="cancelled", message="Cancelled by user.",
                       warning_items=structured_warnings,
                       failure_items=failure_items,
                       skipped_inputs=exact_skipped,
                       failed_inputs=exact_failed)
    except ValueError as exc:
        return _result(status="error", message=str(exc),
                       warning_items=structured_warnings,
                       failure_items=failure_items,
                       skipped_inputs=exact_skipped,
                       failed_inputs=exact_failed)
    pairing_trace = pairing.pairing_trace
    pairing_quality = pairing.pairing_quality
    pairing_diagnostics = pairing.capped_group_diagnostics
    pairing_capped = bool(pairing_diagnostics)
    # Excel hard limits: openpyxl would raise mid-write past the row cap (losing
    # the partial file) and silently lose columns past the column cap.
    biggest = max(len(union), len(rows_t), len(rows_n)) + 1   # + header row
    # Hidden Med-Wid stages and Comparison state masks extend PHYSICAL sheet
    # widths; visible data/filter geometry remains unchanged.
    n_cols = max(lay.physical_n_cols, lay.spot_physical_n_cols,
                 lay.comparison_physical_n_cols)
    limit_err = excel_limit_error(biggest, n_cols)
    if limit_err:
        return _result(status="error", message=limit_err)

    try:
        counts = count_diffs(
            sc, rows_t, rows_n, keys_t, keys_n, union, has_route,
            events.is_cancelled)
    except RunCancelled:
        return _result(status="cancelled", message="Cancelled by user.",
                       warning_items=structured_warnings,
                       failure_items=failure_items,
                       skipped_inputs=exact_skipped,
                       failed_inputs=exact_failed)
    computed_counts = counts

    # Shared model for every output flavor (parsed/aligned/counted ONCE).
    spot_row = counts.get("first_diff_row") or 2
    all_routes = r_both = r_t_only = r_n_only = None
    try:
        if has_route:
            all_routes, r_both, r_t_only, r_n_only = route_coverage(
                keys_t, keys_n, events.is_cancelled)
        derived = _post_count_derivations(
            keys_t, keys_n, union, helper_tokens, rows_t, rows_n,
            include_value_maps=("values" in modes),
            is_cancelled=events.is_cancelled)
        events.on_log(
            f"{a} rows: {len(rows_t):,}   {b} rows: {len(rows_n):,}   "
            f"union: {len(union):,} {sc.id_noun_plural}"
            + (f" across {len(all_routes)} routes" if has_route else ""))
        _raise_if_cancelled(events.is_cancelled)
    except RunCancelled:
        return _result(status="cancelled", message="Cancelled by user.",
                       warning_items=structured_warnings,
                       failure_items=failure_items,
                       skipped_inputs=exact_skipped,
                       failed_inputs=exact_failed)
    only_t, only_n = derived["only_t"], derived["only_n"]
    cmp_rows_t, cmp_rows_n = derived["cmp_rows_t"], derived["cmp_rows_n"]
    hk_t, hk_n = derived["hk_t"], derived["hk_n"]

    cancelled = _result(status="cancelled", message="Cancelled by user.",
                        raw_counts=counts, warning_items=structured_warnings,
                        failure_items=failure_items,
                        skipped_inputs=exact_skipped,
                        failed_inputs=exact_failed)
    for m in modes:
        path = out_paths[m]
        if not output_allowed(path.parent) or not output_allowed(path):
            return ownership_error()
        if m == "values":
            # Everything the formulas would display, precomputed: the writers
            # emit literal results instead of formulas (links/CF/Spot Check/
            # SELF-CHECK stay live). Same _xl_trim/_medwid_norm mirror as the
            # run summary, so the two flavors can never disagree.
            vals = {
                "off": 1 if has_route else 0,
                "by_t": derived["by_t"],
                "by_n": derived["by_n"],
                "row_t": derived["row_t"],
                "row_n": derived["row_n"],
                "routes_t": derived["routes_t"],
                "routes_n": derived["routes_n"],
                "counts": counts,
                "state_masks": counts["state_masks"],
                "n_t": len(rows_t), "n_n": len(rows_n),
                "r_both": len(r_both) if has_route else 0,
                "r_t_only": len(r_t_only) if has_route else 0,
                "r_n_only": len(r_n_only) if has_route else 0,
            }
            events.on_log(f"Writing the VALUES workbook: {path.name}")
        else:
            # The live-formulas flavor now ALSO writes literal lookup keys (the
            # same ones the values flavor uses), not a live COUNTIFS occurrence.
            # A blank key field made the COUNTIFS mis-number occurrences (Excel's
            # blank-criterion quirk), so those rows' lookups failed and the
            # SELF-CHECK read CHECK. Literal keys always match the Comparison
            # sheet's literal occurrence #, and the row universe is build-time
            # static by design anyway (only the field VALUES are live).
            vals = None
            events.on_log(f"Writing the live-formulas workbook: {path.name}")

        # Streaming workbook (see the note above _styled): sheets are created
        # in display order; Summary first so it's the active sheet on open.
        wb = Workbook(write_only=True)
        if has_route and m == "formulas":
            # ~2M live formulas: in automatic mode Excel would recalculate
            # for minutes on open AND after every edit. Ship the workbook in
            # MANUAL calculation mode instead — it opens instantly showing
            # blanks/zeros, the user presses F9 once (the one unavoidable big
            # calc), saves, and from then on opens are instant and edits
            # don't hang. calcOnSave off so saving doesn't sneak the big calc
            # back in. (Per-route files and the VALUES copy stay automatic.)
            wb.calculation.calcMode = "manual"
            wb.calculation.calcOnSave = False
            wb.calculation.fullCalcOnLoad = False
        _write_summary(
            wb, name_a, name_b, len(union), lay, vals=vals,
            warnings=warnings, pairing_diagnostics=pairing_diagnostics,
            source_row_counts=(len(rows_t), len(rows_n)))
        _write_spot_check(wb, lay, len(union), spot_row, union[spot_row - 2],
                          manual_calc=(has_route and m == "formulas"))
        if _write_comparison(
                wb, union, lay, events, vals=vals,
                helper_tokens=helper_tokens) is None:
            return cancelled
        if has_route:
            _write_routes(wb, all_routes, lay, vals=vals)
        # The one-sided rows again, on their own tabs (union order): the rows
        # of routes the other system lacks entirely plus the rows missing
        # from the other side within shared routes.
        if _write_only_sheet(
                wb, a, b, _TAB_COLORS["only_a"], only_t, lay,
                events, vals=vals, helper_tokens=helper_tokens) is None:
            return cancelled
        if _write_only_sheet(
                wb, b, a, _TAB_COLORS["only_b"], only_n, lay,
                events, vals=vals, helper_tokens=helper_tokens) is None:
            return cancelled
        if _write_data_sheet(wb, a, _TAB_COLORS["side_a"], rows_t, lay, events,
                             cmp_rows_t, helper_keys=hk_t,
                             live_medwid_helpers=(m == "formulas")) is None:
            return cancelled
        if _write_data_sheet(wb, b, _TAB_COLORS["side_b"], rows_n, lay, events,
                             cmp_rows_n, helper_keys=hk_n,
                             live_medwid_helpers=(m == "formulas")) is None:
            return cancelled
        if sc.legend_writer is not None:     # append a column-Legend sheet (HL)
            sc.legend_writer(wb)
        if sc.extra_sheet_writer is not None:  # append a familiar-layout rollup sheet
            sc.extra_sheet_writer(wb, {"rows_a": rows_t, "rows_b": rows_n,
                                       "has_route": has_route, "sc": sc,
                                       "side_a": name_a, "side_b": name_b,
                                       # F1 (additive): the pairing run_compare already
                                       # computed, so a big rollup needn't redo it.
                                       # Writers may ignore these; output unaffected.
                                        "keys_a": keys_t, "keys_b": keys_n,
                                        "union": union,
                                        # E1: compact, typed comparison truth in
                                        # displayed-field order. Familiar-layout
                                        # writers may consume it without parsing
                                        # presentation text; existing writers may
                                        # ignore the additive context keys.
                                        "state_masks": counts["state_masks"],
                                        "state_field_indices": tuple(lay.field_indices),
                                        # events lets a big rollup report progress so it
                                        # isn't a multi-minute silent gap before "Saving…".
                                       "events": events})
        if _write_snapshot_sheet(
                wb, a, rows_t, lay, events, hk_t) is None:
            return cancelled
        if _write_snapshot_sheet(
                wb, b, rows_n, lay, events, hk_n) is None:
            return cancelled
        if provenance is not None:           # CMP-AUD-076 (opt-in, additive)
            _write_provenance_sheet(wb, provenance)
        events.on_log("Saving…")
        if not output_allowed(path.parent) or not output_allowed(path):
            close_unsaved(wb)
            return ownership_error()
        path.parent.mkdir(parents=True, exist_ok=True)
        if not output_allowed(path.parent) or not output_allowed(path):
            close_unsaved(wb)
            return ownership_error()
        try:
            wb.save(path)
        except PermissionError:
            return _result(
                status="error",
                message=(f"Could not save {path.name}.\n\n"
                         "The file is probably open in Excel. Close it and try again."),
                raw_counts=counts,
                warning_items=structured_warnings,
                failure_items=failure_items,
                skipped_inputs=exact_skipped,
                failed_inputs=exact_failed)
        if not output_allowed(path):
            return ownership_error()

    def _route_list(label, routes):
        shown = ", ".join(routes[:15]) + (", …" if len(routes) > 15 else "")
        return f"{label} ({len(routes)}): {shown}" if routes else f"{label}: none"

    # The quick answer first: most comparisons exist to confirm "nothing
    # changed", so say so in one loud line (mirrored by the workbook's
    # Summary verdict; the GUI also keys its result dialog on `verdict`).
    one_sided = counts["t_only"] + counts["n_only"]
    matches = counts["diff_cells"] == 0 and one_sided == 0
    # Unreadable inputs or capped duplicate identity make the comparison
    # INCOMPLETE: neither condition may certify a clean match.
    incomplete = input_incomplete or pairing_capped
    if pairing_capped:
        qualifier = (" and incomplete input coverage" if input_incomplete else "")
        verdict_line = (
            "⚠ PARTIAL / PAIRING LIMIT — positional fallback observed "
            f"{counts['diff_cells']:,} differing cell(s) on "
            f"{counts['diff_rows']:,} matched row(s), "
            f"{counts['t_only'] + counts['n_only']:,} one-sided row(s)"
            f"{qualifier}. These are diagnostic counts, not certified "
            "differences; re-scope and regenerate.")
    elif matches and not incomplete:
        verdict_line = (f"✓ EVERYTHING MATCHES — all {len(union):,} "
                        f"{sc.id_noun_plural} are identical in both "
                        f"{sc.sides_noun}.")
    elif matches and input_incomplete:
        verdict_line = (f"⚠ COULD NOT COMPARE EVERYTHING — {len(warnings)} input "
                        f"file(s) were unreadable and skipped; the {len(union):,} "
                        f"{sc.id_noun_plural} that WERE compared all match.")
    else:
        verdict_line = (f"✗ DIFFERENCES FOUND — {counts['diff_cells']:,} "
                        f"differing cell(s) on {counts['diff_rows']:,} "
                        f"matched row(s); {counts['t_only']:,} row(s) only "
                        f"in {a}, {counts['n_only']:,} only in {b}.")

    Nouns = sc.id_noun_plural.capitalize()
    lines = [
        verdict_line,
        f"{Nouns} in both {sc.sides_noun}:   {counts['both']:,}",
        f"In {a} only / in {b} only: {counts['t_only']:,} / {counts['n_only']:,} "
        "(each listed on its own 'Only in …' sheet)",
        f"Matched rows with differences: {counts['diff_rows']:,} "
        f"({counts['diff_cells']:,} differing cells); "
        f"{counts['identical']:,} fully identical",
    ]
    if has_route:
        lines += [
            f"Routes covered by both {sc.sides_noun}: {len(r_both)}",
            _route_list(f"Routes only in {a} (missing from {b})", r_t_only),
            _route_list(f"Routes only in {b} (missing from {a})", r_n_only),
            "Those routes' rows are included — tinted 'entire route' on the "
            "'Only in …' sheets; per-route breakdown on the Routes sheet.",
        ]
        if "formulas" in modes:
            lines.append(
                "Note: the live-formulas workbook opens in MANUAL calculation "
                "— press F9 in Excel to calculate (first time takes a few "
                "minutes), then save. The Summary's SELF-CHECK rows should "
                "all read OK." + ("  The values copy opens ready — nothing "
                                  "to calculate." if "values" in modes else ""))
    if input_incomplete:
        lines.append(f"⚠ {len(warnings)} input file(s) skipped (unreadable) and "
                     "left OUT of this comparison — re-export/repair and re-run:")
        for w in warnings[:20]:
            lines.append(f"    – {w}")
        if len(warnings) > 20:
            lines.append(f"    – …and {len(warnings) - 20} more (see the log).")
    if pairing_capped:
        lines.append(
            f"⚠ {len(pairing_diagnostics)} duplicate key group(s) exceeded "
            "the exact-pairing cap. Displayed differences are under "
            "deterministic positional fallback, not a certified optimum; "
            "re-scope and re-run before trusting a clean result.")
        for diagnostic in pairing_diagnostics[:20]:
            key = " / ".join(diagnostic.key_components)
            lines.append(
                f"    – {key}: {diagnostic.side_a_size} × "
                f"{diagnostic.side_b_size}; fallback cost "
                f"{diagnostic.fallback_cost}")
        if len(pairing_diagnostics) > 20:
            lines.append(
                f"    – …and {len(pairing_diagnostics) - 20} more "
                "(see the typed outcome sidecar).")
    if "formulas" in modes:
        lines.append(f"Live-formulas file: {out_paths['formulas']}")
    if "values" in modes:
        lines.append(f"Values file: {out_paths['values']}")
    primary = out_paths["formulas" if "formulas" in modes else "values"]
    # Incomplete ⇒ never certify a clean match (GUI greens only on "match").
    verdict = "diff" if (incomplete or not matches) else "match"
    return _result(
        status="ok",
        output_path=str(primary),
        summary_lines=lines,
        verdict=verdict,
        completion=outcome.PARTIAL if incomplete else outcome.COMPLETE,
        raw_counts=counts,
        warning_items=structured_warnings,
        failure_items=failure_items,
        skipped_inputs=exact_skipped,
        failed_inputs=exact_failed,
    )
