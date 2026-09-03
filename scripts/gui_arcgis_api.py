"""GuiArcgisMixin — the ArcGIS tab's endpoints (v0.29.0; reports sub-tab
v0.39.0; the "Reports vs layers" MATRIX + library 2026-09-02).

The tab owns everything built FROM the ArcGIS layer library, in two sub-tabs:

  * **Reports vs layers** (the main view) — every TSMIS report rendered from
    the layers and diffed against the app's own export of it, as a by-day
    MATRIX: rows = the `arcgis_reports` registry (the two rendered so far plus
    the rows still waiting on a build, greyed), columns = exported days, each
    cell = the report's SINGLE layer build vs that day's consolidated export.
    The library card above it names the staged drop (export date + content
    fingerprint) and each row header says whether its build is from that drop.
    Unlike vs-TSN this is TSMIS vs TSMIS, so the two sides SHOULD agree — which
    is why each build's as-of date sits beside the export day rather than hiding
    a mismatch that would make the comparison measure network change.
    Engine: `arcgis_matrix`; jobs ride the shared matrix queue (`which:"arcgis"`
    compares, `kind:"arcgis_build"` builds).
  * **Clean Road vs TSN** — the layer-library stock vs the 40-layer manifest,
    the CA HIGHWAYS overlay build (`consolidate_clean_highway`), and the
    ArcGIS-vs-TSN comparison (`compare_clean_highway_tsn`). Intersections and
    Ramps surface as staged skeletons only.

The Clean Road comparison rides the shared `_begin_compare` claim → save-dialog
→ launch tail like every other manual comparison. Composition only — every
`self._*` it touches lives on GuiApi.
"""
import logging
from pathlib import Path

import settings
from gui_endpoint import _api_method
from gui_worker import ConsolidateWorker
from gui_worker_matrix import ArcgisMatrixCompareWorker, ArcgisReportBuildWorker

ui_log = logging.getLogger("tsmis.ui")


class GuiArcgisMixin:
    # ------------------------------------------------------------------ #
    # Reports vs layers — the matrix + the library (2026-09-02)
    # ------------------------------------------------------------------ #
    def _arcgis_matrix_snapshot(self):
        """The Reports-vs-layers snapshot with the user's source / day columns /
        hidden rows / drag order applied."""
        import arcgis_matrix

        return arcgis_matrix.ag_matrix_snapshot(
            settings.get_arcgis_matrix_source(), settings.get_arcgis_matrix_days(),
            hidden=settings.get_arcgis_matrix_hidden(),
            row_order=settings.get_arcgis_matrix_row_order())

    @staticmethod
    def _arcgis_row_label(row_key):
        import arcgis_reports
        return arcgis_reports.label_of(row_key)

    def _arcgis_job_label(self, scope, row=None, date=None):
        rl = self._arcgis_row_label(row) if row else None
        if scope == "cell":
            return f"Rebuild {rl} — {date} vs layers"
        if scope == "row":
            return f"Rebuild {rl} — all days vs layers"
        if scope == "column":
            return f"Rebuild all reports — {date} vs layers"
        if scope == "stale":
            return "Refresh stale Reports-vs-layers comparisons"
        return "Rebuild all Reports-vs-layers comparisons"

    @_api_method
    def arcgis_matrix_info(self):
        """The Reports-vs-layers snapshot (incl. the library) + the add-day
        picker's available days."""
        import arcgis_matrix

        snap = self._arcgis_matrix_snapshot()
        snap["available_days"] = arcgis_matrix.available_days(snap["source"])
        snap["available_day_reports"] = arcgis_matrix.available_day_reports(
            snap["source"])
        self._push_state()
        return snap

    @_api_method
    def set_arcgis_matrix_source(self, source):
        """Set the matrix data source (day columns are dates within it)."""
        import arcgis_matrix
        import matrix

        if source not in arcgis_matrix.sources():
            return {"error": "Unknown data source."}
        settings.set_arcgis_matrix_source(source)
        avail = set(arcgis_matrix.available_days(source))
        kept = [d for d in settings.get_arcgis_matrix_days() if d in avail]
        if kept != settings.get_arcgis_matrix_days():
            settings.set_arcgis_matrix_days(kept)
        self._emit_log("Reports-vs-layers matrix source set to "
                       f"{matrix.default_env_label(source)}.")
        self._push_state()
        return {"ok": True, "source": source, "days": kept}

    @_api_method
    def add_arcgis_matrix_day(self, date):
        """Add a day COLUMN (a date with an export of some report for the source)."""
        import arcgis_matrix

        if date not in arcgis_matrix.available_days(settings.get_arcgis_matrix_source()):
            return {"error": "That day has no export for this source (only exported "
                             "days can be added)."}
        days = settings.get_arcgis_matrix_days()
        if date not in days:
            settings.set_arcgis_matrix_days(days + [date])
        self._push_state()
        return {"ok": True, "days": settings.get_arcgis_matrix_days()}

    @_api_method
    def remove_arcgis_matrix_day(self, date):
        settings.set_arcgis_matrix_days(
            [d for d in settings.get_arcgis_matrix_days() if d != date])
        self._push_state()
        return {"ok": True, "days": settings.get_arcgis_matrix_days()}

    @_api_method
    def set_arcgis_matrix_report(self, row_key, visible):
        """Show/hide a report ROW. At least one stays on."""
        import arcgis_matrix

        keys = arcgis_matrix.row_keys()
        if row_key not in keys:
            return {"error": "Unknown report for the matrix."}
        hidden = set(settings.get_arcgis_matrix_hidden())
        if visible:
            hidden.discard(row_key)
        else:
            hidden.add(row_key)
        if len(hidden & keys) >= len(keys):
            return {"error": "Keep at least one report on the matrix."}
        settings.set_arcgis_matrix_hidden(sorted(hidden))
        self._push_state()
        return {"ok": True, "hidden": sorted(hidden)}

    @_api_method
    def set_arcgis_matrix_row_order(self, keys):
        import arcgis_matrix

        valid = arcgis_matrix.row_keys()
        clean = [k for k in (keys or []) if isinstance(k, str) and k in valid]
        settings.set_arcgis_matrix_row_order(clean)
        self._push_state()
        return {"ok": True, "order": clean}

    @_api_method
    def set_arcgis_matrix_day_order(self, days):
        """Persist the drag-to-reorder DAY-column order. Any current day the
        client omitted is appended in its existing order (a reorder never
        drops a column)."""
        current = settings.get_arcgis_matrix_days()
        clean = [d for d in (days or []) if isinstance(d, str) and d in current]
        clean += [d for d in current if d not in clean]
        settings.set_arcgis_matrix_days(clean)
        self._push_state()
        return {"ok": True, "days": clean}

    @_api_method
    def set_arcgis_matrix_formulas(self, on):
        settings.set_arcgis_matrix_formulas(bool(on))
        self._push_state()
        return {"ok": True, "on": bool(on)}

    @_api_method
    def build_arcgis_matrix_cell(self, row_key, date):
        """Queue a (re)build of ONE (report, day) comparison."""
        import matrix

        snap = self._arcgis_matrix_snapshot()
        if row_key not in {r["key"] for r in snap["all_rows"]}:
            return {"error": "Unknown report for the matrix."}
        if date not in snap["days"]:
            return {"error": "Add that day first."}
        cmp = snap["cells"].get(row_key, {}).get(date, {}).get("cmp")
        reason = matrix.cell_unbuildable_reason(cmp)   # CMP-AUD-103
        if reason:
            return {"error": reason}
        job = self._make_job("compare", "cell",
                             self._arcgis_job_label("cell", row_key, date),
                             row=row_key, env=date, which="arcgis")
        return self._enqueue_matrix_job(job)

    @_api_method
    def rebuild_arcgis_matrix(self, scope="stale", row=None, date=None, force=False):
        """Queue a rebuild in scope ('stale'/'all'), optionally scoped to one
        report row or one day column. `force` also rebuilds the day's persistent
        consolidated export. {nothing:True} when idle with nothing to do."""
        import arcgis_matrix

        snap = self._arcgis_matrix_snapshot()
        scope = scope if scope in ("stale", "all") else "stale"
        if row and row not in {r["key"] for r in snap["all_rows"]}:
            return {"error": "Unknown report row for the rebuild."}
        if date and date not in snap["days"]:
            return {"error": "Unknown day for the rebuild."}
        row = row or None
        date = date or None
        job_scope = "row" if row else "column" if date else scope
        with self._lock:
            idle = not self._task and not self._queue
        if idle:
            probe_scope = "all" if force else scope
            cells = arcgis_matrix.cells_to_rebuild(snap, scope=probe_scope,
                                                   row=row, date=date)
            if not cells:
                return {"ok": True, "nothing": True}
        job = self._make_job("compare", job_scope,
                             self._arcgis_job_label(job_scope, row, date),
                             row=row, env=date, which="arcgis", force=force)
        return self._enqueue_matrix_job(job)

    @_api_method
    def build_arcgis_report(self, row_key, asof=None):
        """Queue the (re)build of a report's SINGLE layer build. `asof` is the
        reconstruction date; empty means the staged drop's own export date. Rides
        the matrix queue so it lines up behind (or ahead of) the comparisons that
        need it, and is cancellable the same way."""
        import arcgis_reports

        if not arcgis_reports.is_report(row_key):
            return {"error": "Unknown report."}
        if not arcgis_reports.can_build(row_key):
            return {"error": f"{arcgis_reports.label_of(row_key)} cannot be built "
                             "from the layers yet."}
        asof = (asof or "").strip()
        if asof:
            import clean_road_layers as crl
            if crl.to_serial(asof) is None:
                return {"error": f"Not a usable as-of date: {asof!r} (use YYYY-MM-DD)."}
        label = (f"Build {arcgis_reports.label_of(row_key)} from the layers"
                 + (f" as of {asof}" if asof else ""))
        job = self._make_job("arcgis_build", "cell", label, row=row_key,
                             which="arcgis", asof=asof or None)
        return self._enqueue_matrix_job(job)

    @_api_method
    def open_arcgis_report(self, row_key):
        """Open a report's built workbook."""
        import arcgis_matrix
        import arcgis_reports

        if not arcgis_reports.is_report(row_key):
            return {"error": "Unknown report."}
        bs = arcgis_matrix.build_state(row_key)
        if not bs.get("built"):
            return {"error": "This report hasn't been built from the layers yet."}
        self._open_file(Path(bs["path"]))
        return {"ok": True}

    @_api_method
    def open_arcgis_cell_comparison(self, row_key, date):
        """Open ONE Reports-vs-layers comparison VALUES workbook."""
        import arcgis_matrix

        snap = self._arcgis_matrix_snapshot()
        if date not in snap["days"] or row_key not in snap["rows"]:
            return {"error": "Unknown cell."}
        path = arcgis_matrix.day_out_path(date, snap["source"], row_key)
        if not path.exists():
            return {"error": "No comparison built yet — use “⟳ rebuild” first."}
        self._open_file(path)
        return {"ok": True}

    @_api_method
    def open_arcgis_comparisons_folder(self):
        """Open the Reports-vs-layers comparison store (comparisons/arcgis-by-day/)."""
        import arcgis_matrix

        self._open_folder(arcgis_matrix.ag_root())
        return {"ok": True}

    @_api_method
    def open_arcgis_reports_folder(self):
        import arcgis_report_highway_detail as ah
        ah.OUT_DIR.mkdir(parents=True, exist_ok=True)
        self._open_folder(ah.OUT_DIR)
        return {"ok": True}

    # -- the queue's dispatch hooks (called by gui_matrix's dispatcher) -------- #
    def _resolve_arcgis_cells(self, job):
        """[(date, row)] for a Reports-vs-layers compare job. 'cell' = the one
        explicit cell; row/column/all/stale defer to the staleness-aware list."""
        import arcgis_matrix
        import matrix

        snap = self._arcgis_matrix_snapshot()
        scope = job["scope"]
        if scope == "cell":
            row, date = job["row"], job["env"]
            if row not in snap["rows"] or date not in snap["days"]:
                return []
            cmp = snap["cells"].get(row, {}).get(date, {}).get("cmp")   # CMP-AUD-103
            if not matrix.cell_buildable(cmp):
                return []
            return [(date, row)]
        rebuild_scope = "stale" if scope == "stale" else "all"
        return arcgis_matrix.cells_to_rebuild(snap, scope=rebuild_scope,
                                              row=job.get("row"), date=job.get("env"))

    def _dispatch_arcgis_compare_job(self, job):
        source = settings.get_arcgis_matrix_source()
        cells = self._resolve_arcgis_cells(job)
        if not cells:
            return False
        with self._lock:
            self._matrix = {"phase": "comparing", "row": job.get("row"),
                            "cell": job.get("env"), "done": 0, "total": len(cells)}
        self._emit_log(f"{job['label']} — {len(cells)} Reports-vs-layers comparison(s)…")
        self._set_dot("busy", "Comparing…")
        self._emit({"t": "run_started", "mode": "consolidate", "label": "Comparing…",
                    "workers": 1})
        ArcgisMatrixCompareWorker(
            source, cells, self._gated_queue(), self.cancel_event,
            force_consolidate=job.get("force", False),
            also_formulas=settings.get_arcgis_matrix_formulas()).start()
        return True

    def _dispatch_arcgis_build_job(self, job):
        row = job.get("row")
        with self._lock:
            self._matrix = {"phase": "building", "row": row, "cell": None,
                            "done": 0, "total": 1}
        self._emit_log(f"{job['label']}…")
        self._set_dot("busy", "Building the report from the layers…")
        self._emit({"t": "run_started", "mode": "consolidate",
                    "label": job["label"] + "…", "workers": 1})
        ArcgisReportBuildWorker(row, job.get("asof"), self._gated_queue(),
                                self.cancel_event).start()
        return True

    # ------------------------------------------------------------------ #
    # Clean Road vs TSN (v0.29.0)
    # ------------------------------------------------------------------ #
    @_api_method
    def arcgis_status(self):
        """The Clean Road sub-tab's one status payload: library stock vs the
        40-layer manifest, the CA HIGHWAYS build/TSN readiness, and the staged
        skeletons. Pure filesystem; no task lock, no browser."""
        import clean_road_layers as crl
        import consolidate_clean_highway as cch
        import consolidation_meta
        import tsn_library

        lib = crl.inventory()
        highway_missing = [n for n in cch.HIGHWAY_LAYERS
                           if n not in lib["present"]]
        built = cch.OUT_PATH
        built_info = {"exists": built.is_file(), "path": str(built)}
        if built_info["exists"]:
            record = consolidation_meta.read_outcome(built)
            if record is not None and record.current:
                built_info["completion"] = record.completion
            built_info.update(self._arcgis_built_marker(built))
        default_asof = None
        tsn_raw = False
        try:
            raw_root = Path(tsn_library.raw_dir("clean_highway"))
            tsn_raw = any(p.is_file() and not p.name.startswith("~$")
                          for p in raw_root.glob("*.xlsx"))
            if tsn_raw:
                default_asof = cch.resolve_default_asof().isoformat()
        except (OSError, ValueError) as e:
            ui_log.info("arcgis: default as-of unavailable (%s: %s)",
                        type(e).__name__, e)
        return {
            "root": str(crl.root()),
            "expected": len(crl.EXPECTED_LAYERS),
            "staged": len(lib["present"]),
            "missing": lib["missing"],
            "unknown": lib["unknown"],
            "index_present": lib["index"] is not None,
            "highway": {
                "layers_ok": not highway_missing,
                "missing": highway_missing,
                "built": built_info,
                "tsn_raw": tsn_raw,
                "default_asof": default_asof,
            },
        }

    @staticmethod
    def _arcgis_built_marker(path):
        """The built workbook's own as-of/build facts from its marker sheet
        (absent/unreadable facts read as unknown, never invented)."""
        import clean_highway_columns as chc
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                if chc.ARC_MARKER_SHEET not in wb.sheetnames:
                    return {}
                facts = {}
                for r in wb[chc.ARC_MARKER_SHEET].iter_rows(values_only=True):
                    if not r or r[0] is None:
                        continue
                    key = str(r[0]).strip()
                    if key == "As-of date":
                        facts["asof"] = str(r[1])
                    elif key == "Build version":
                        facts["build_version"] = r[1]
                return facts
            finally:
                wb.close()
        except Exception as e:      # silent-ok: a status card probe; an unreadable marker shows as unknown, the build/compare paths still gate hard
            ui_log.info("arcgis: built marker unreadable (%s: %s)",
                        type(e).__name__, e)
            return {}

    @_api_method
    def open_arcgis_layers_folder(self):
        import arcgis_layers
        arcgis_layers.ensure_layout()
        self._open_folder(arcgis_layers.root())
        return {"ok": True}

    @_api_method
    def open_arcgis_output_folder(self):
        import consolidate_clean_highway as cch
        cch.OUT_DIR.mkdir(parents=True, exist_ok=True)
        self._open_folder(cch.OUT_DIR)
        return {"ok": True}

    @_api_method
    def start_arcgis_build(self, asof=None):
        """Build the CA HIGHWAYS clean-road workbook from the layer library.
        `asof` is an optional ISO date; empty resolves to the staged TSN
        extract's own date. The destination is app-owned
        (output/arcgis_cleanroad) and a rebuild replaces it by design."""
        import consolidate_clean_highway as cch

        asof = (asof or "").strip() or None
        err = self._claim_task_error("consolidate")
        if err:
            return err
        self.cancel_event.clear()
        label = "Clean Road Highway (ArcGIS)"
        self._emit_log(f"Starting ArcGIS build: {label}"
                       + (f"   ·   as of {asof}" if asof else ""))
        self._set_dot("busy", "Building from ArcGIS layers…")
        self._emit({"t": "run_started", "mode": "consolidate",
                    "label": "Building Clean Road Highway from the ArcGIS "
                             "layers…"})
        self._push_state()

        def build(events=None, confirm_overwrite=None, day=None):
            return cch.consolidate(events=events,
                                   confirm_overwrite=confirm_overwrite,
                                   day=day, asof=asof)

        ConsolidateWorker(build, self._gated_queue(), self.cancel_event,
                          lambda _p: True).start()
        return {"ok": True}

    @_api_method
    def start_arcgis_compare(self, want_formulas=True, want_values=True):
        """Compare the built CA HIGHWAYS workbook against the TSN clean-road
        extract. The TSN library normalization builds (or reuses) inside the
        worker; the save dialog owns the destination + overwrite question."""
        import compare_clean_highway_tsn as cht
        import consolidate_clean_highway as cch
        import paths
        import tsn_library

        built = cch.OUT_PATH
        if not built.is_file():
            return {"error": "Build the Clean Road Highway workbook first — "
                             "the comparison reads it as the ArcGIS side."}
        try:
            raw_root = Path(tsn_library.raw_dir("clean_highway"))
            has_raw = any(p.is_file() and not p.name.startswith("~$")
                          for p in raw_root.glob("*.xlsx"))
        except OSError:  # silent-ok: a pure presence probe — an unreadable raw folder reads as not-staged and the endpoint returns the stage-it-first message
            has_raw = False
        if not has_raw:
            return {"error": "Stage the TSN CA HIGHWAYS extract in the TSN "
                             "library first (Settings → TSN reports → Clean "
                             "Road Highway)."}
        mode = self._compare_mode(want_formulas, want_values)
        if mode is None:
            return {"error": "Tick at least one output (values and/or live "
                             "formulas)."}
        tsn_path = Path(tsn_library.consolidated_path("clean_highway"))

        def build(out_path, events=None, confirm_overwrite=None, day=None):
            res = tsn_library.build_consolidated(
                "clean_highway", events=events, confirm_overwrite=lambda p: True)
            if res.status != "ok":
                return res
            return cht.compare(built, tsn_path, out_path, events=events,
                               confirm_overwrite=confirm_overwrite, mode=mode)

        sources = tuple(p for p in (built, tsn_path) if p.is_file())
        # M2-E: auto-save to the standardized output/comparisons/arcgis/ (beside the
        # manual folder) instead of the ArcGIS build output dir; "Save elsewhere…" stays.
        return self._begin_compare(
            "Clean Road Highway — ArcGIS vs TSN", mode, paths.arcgis_comparisons_dir(),
            lambda: cht.suggest_name(built), build, source_paths=sources)
