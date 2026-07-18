"""Cross-environment comparison: the same report exported from two different
data source / environment combinations (e.g. SSOR-prod vs ARS-prod, or
SSOR-prod today vs SSOR-prod last month), compared cell-for-cell.

Inputs are two RUN FOLDERS (output/<YYYY-MM-DD src-env>/ — or a report
subfolder picked directly). No consolidation step is needed first: the
per-route files of the chosen report are read straight from both folders
and merged in memory exactly the way the consolidators would (Route column
prepended, header locked from the first readable file), then handed to the
proven compare_core engine — so the output is the same approved discrepancy
workbook the TSMIS-vs-TSN comparison produces, with the environment names
("SSOR-PROD", "ARS-DEV", …) as the two sides.

One adapter per report type (REPORTS / the per-report constants below):
  * Ramp Detail / Highway Sequence / Highway Log — per-route XLSX exports,
    compared in the consolidated shape (Route + the report's own columns;
    the column layout is locked from the files and must match between the
    two folders). Highway Log keeps its Med Wid zero-pad normalization.
  * Ramp Summary — per-route PDFs, parsed with the consolidator's own parser
    (consolidate_ramp_summary.parse_pdf) into one row per route, compared
    route-by-route (the route is the row key).

Console-free, same contract as the other comparison modules: progress via
events.on_log, overwrite via confirm_overwrite, cancel honored per file and
inside the engine, ConsolidateResult returned. The GUI's Compare tab drives
this through the COMPARE_REPORTS registry ("folders" input kind).
"""
import logging
import re
import shutil
import tempfile
from dataclasses import replace
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False

# Route-key normalizer: the Ramp Summary route can come from the PDF title
# (unpadded, "1") or the filename ("001"); zero-pad the numeric part so the two
# sides key consistently and a route doesn't split into two one-sided rows.
_ROUTE_KEY_RE = re.compile(r"^(\d+)(\w*)$")


def _norm_route_key(token):
    m = _ROUTE_KEY_RE.match(str(token).strip())
    return m.group(1).zfill(3) + m.group(2).upper() if m else str(token).strip().upper()

import artifact_store
import compare_highway_log as _hl
import compare_tsn_common as ctc
import consolidate_intersection_summary as _is
import consolidate_ramp_summary as _rs
from compare_core import CompareSchema, normalize_value, run_compare
from compare_tsn_common import row_has_data


def _member_census(files):
    """[{name, size, mtime_ns}] for the exact discovered member set, statted
    before any loader reads (CMP-AUD-076 folder-kind provenance). Raises
    ValueError when a discovered member can't be statted — an input we cannot
    identify must not be compared."""
    census = []
    for f in files:
        p = Path(f)
        try:
            st = p.stat()
        except OSError as e:
            raise ValueError(f"Could not capture the input identity for "
                             f"{p.name}: {type(e).__name__}: {e}")
        census.append({"name": p.name, "size": st.st_size,
                       "mtime_ns": st.st_mtime_ns})
    return census
from comparison_contract import (ComparisonCounts, ComparisonOutcome, LoadedSide,
                                 comparison_result_boundary)
from events import ConsolidateResult, Events
from paths import OUTPUT_ROOT, parse_run_folder, today_str

log = logging.getLogger("tsmis.compare")

# Pull the route token out of "<prefix>_route_<ROUTE>.<ext>" (same rule as
# consolidate_xlsx_base, extended to PDFs for the Ramp Summary).
_ROUTE_FROM_NAME = re.compile(r"_route_(\w+)\.(?:xlsx|pdf)$", re.IGNORECASE)


def _route_from_name(path):
    m = _ROUTE_FROM_NAME.search(path.name)
    return m.group(1).upper() if m else path.stem


def side_label(folder):
    """A short side name for one input folder, used as that side's sheet/tab
    name and in every label: run folders become "SSOR-PROD" style; anything
    else falls back to a sanitized folder name."""
    folder = Path(folder)
    parsed = parse_run_folder(folder.name)
    if parsed is None and folder.parent != folder:
        parsed = parse_run_folder(folder.parent.name)   # report subdir picked
    if parsed:
        _day, src, env = parsed
        return f"{src}-{env}".upper()
    clean = re.sub(r"[\[\]\*\?:/\\']+", " ", folder.name).strip()
    return (clean or "FOLDER")[:20]


# Excel caps sheet names at 31 chars; the longest sheet we derive from a side
# label is "Only in <label>" (8 + label), so a side label must fit in 23.
_SIDE_LABEL_CAP = 31 - len("Only in ")          # = 23


def _cap_label(s, limit=_SIDE_LABEL_CAP):
    """Cap a derived side label to `limit` chars WITHOUT dropping its trailing
    distinguisher. The labels built below end in the part that keeps two
    same-source sides apart -- a run date (" 2026-06-11") or an " (A)"/" (B)"
    suffix. A plain end-truncation (s[:limit]) would cut exactly that, collapsing
    two distinct sides into the same prefix (then the A/B fallback fires and the
    real provenance is lost). Trim the BASE and keep the suffix instead."""
    if len(s) <= limit:
        return s
    m = re.search(r"(?: \d{4}-\d{2}-\d{2}| \([AB]\))$", s)
    if not m:
        return s[:limit]
    suffix = m.group(0)
    keep = max(0, limit - len(suffix))
    return (s[:m.start()][:keep] + suffix)[:limit]


def _side_labels(dir_a, dir_b):
    """Distinct side names for the two folders. Same src-env on both sides
    (e.g. prod today vs prod last month) gets the run date appended; still
    identical falls back to A/B suffixes (sheet names must differ)."""
    la, lb = side_label(dir_a), side_label(dir_b)
    if la == lb:
        for folder, label in ((dir_a, la), (dir_b, lb)):
            parsed = parse_run_folder(Path(folder).name) \
                or parse_run_folder(Path(folder).parent.name)
            if parsed:
                day = parsed[0]
                if folder is dir_a:
                    la = f"{label} {day}"
                else:
                    lb = f"{label} {day}"
    if la == lb:
        la, lb = f"{la} (A)", f"{lb} (B)"
    # Cap so the longest derived sheet name ("Only in <label>", 8 + label) fits
    # Excel's 31-char limit. _cap_label trims the BASE, not the trailing
    # distinguisher (run date / (A)/(B)), so two same-source sides stay distinct
    # under the cap; fall back to Side A/B only if they still collide.
    la, lb = _cap_label(la), _cap_label(lb)
    if la == lb:
        la, lb = "Side A", "Side B"
    return la, lb


def _find_input_dir(base, subdir, pattern):
    """The folder actually holding the report files: <base>/<subdir> when the
    user picked a run folder, else <base> itself (they browsed straight to a
    report folder). Returns (dir, files) — files possibly empty."""
    base = Path(base)
    for candidate in (base / subdir, base):
        # Exclude Office owner-lock stubs (~$foo.xlsx) — they appear the moment a
        # per-route export is open in Excel and are not report inputs
        # (CMP-AUD-029). Counting one as a member would open-fail and turn
        # identical exports into an incomplete comparison merely because a file
        # was open. Every other XLSX path already filters them (the shared
        # consolidator, Intersection Summary); this is the generic cross-env one.
        files = sorted(p for p in candidate.glob(pattern)
                       if not p.name.startswith("~$")) if candidate.is_dir() else []
        if files:
            return candidate, files
    return base / subdir, []


# ---------------------------------------------------------------------------
# Input loaders
# ---------------------------------------------------------------------------

def _load_xlsx_side(folder, label, subdir, sheet_name, report_name, events,
                    expected_header=None, value_normalizer=None):
    """Read every per-route XLSX under one side into consolidated-shape rows
    ([route, *row]) the way the consolidators do: the header is locked from
    the first readable file (or must equal `expected_header` when the report
    pins one); files that disagree are skipped LOUDLY. Returns
    (rows, header, skipped) — `skipped` is the list of "<side> <file>: <reason>"
    strings that the caller folds into the comparison's incompleteness warning,
    so a route unreadable on a side can never masquerade as a clean match.
    Raises ValueError with a user-safe message when nothing is readable.
    `value_normalizer` (CMP-AUD-047) replaces the generic ``normalize_value``
    with the report's OWN projection so cross-environment loading judges
    values exactly like the report's dedicated comparator (Highway Log's
    tab/newline collapse); None keeps the generic projection."""
    in_dir, files = _find_input_dir(folder, subdir, "*.xlsx")
    if not files:
        raise ValueError(
            f"No {report_name} files were found for the {label} side:\n{in_dir}\n\n"
            f"Export the {report_name} report on that environment first.")
    header = list(expected_header) if expected_header else None
    rows = []
    skipped = []
    seen_routes = set()
    for i, p in enumerate(files, 1):
        if events.is_cancelled():
            raise ValueError("Cancelled by user.")
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
        except Exception as e:
            events.on_log(f"  [{label}] {p.name}: could not open "
                          f"({type(e).__name__}); skipping")
            skipped.append(f"{label} {p.name}: could not open "
                           f"({type(e).__name__})")
            continue
        try:
            if sheet_name not in wb.sheetnames:
                events.on_log(f"  [{label}] {p.name}: sheet '{sheet_name}' "
                              "missing; skipping")
                skipped.append(f"{label} {p.name}: sheet '{sheet_name}' missing")
                continue
            rows_iter = wb[sheet_name].iter_rows(values_only=True)
            h = [v for v in next(rows_iter, [])]
            while h and h[-1] in (None, ""):
                h.pop()
            # Give INTERNAL unnamed columns a stable, identifiable label (the
            # Highway Sequence export has real but header-less columns) so they
            # don't show as blank fields in the Summary/Comparison. Positional,
            # so it's consistent across files (the h != header check still holds).
            h = [v if (v is not None and str(v).strip() != "")
                 else f"(col {get_column_letter(i + 1)})" for i, v in enumerate(h)]
            if header is None:
                header = h
            if h != header:
                events.on_log(f"  [{label}] {p.name}: column layout differs; "
                              "skipping")
                skipped.append(f"{label} {p.name}: column layout differs from "
                               "the other files")
                continue
            # Route identity is the NORMALIZED token from the export's
            # "<report>_route_<token>.xlsx" name (CMP-AUD-031) — zero-pad-normalized
            # so "route_1" and "route_001" are ONE route (not two one-sided rows),
            # and NEVER an arbitrary file stem promoted to a route. A file without
            # that naming contract is not a per-route export and is skipped LOUDLY.
            m_route = _ROUTE_FROM_NAME.search(p.name)
            if not m_route:
                events.on_log(f"  [{label}] {p.name}: not a "
                              "'<report>_route_<n>' export; skipping")
                skipped.append(f"{label} {p.name}: not a recognized "
                               "'..._route_<n>.xlsx' export name")
                continue
            route = _norm_route_key(m_route.group(1))
            # One route per file per side (CMP-AUD-030): a repeat is a stale copy
            # or a split export silently doubling coverage — disclose it as
            # incomplete instead of concatenating the rows.
            if route in seen_routes:
                events.on_log(f"  [{label}] {p.name}: duplicate route {route}; "
                              "skipping")
                skipped.append(f"{label} {p.name}: duplicate route {route} "
                               "(already provided by another file on this side)")
                continue
            seen_routes.add(route)
            n = len(header)
            count = 0
            norm = value_normalizer if value_normalizer is not None \
                else normalize_value
            for r in rows_iter:
                r = list(r)[:n] + [None] * max(0, n - len(r))
                if row_has_data(r):
                    rows.append([route] + [norm(v) for v in r])
                    count += 1
            events.on_log(f"  [{label}] [{i:>3}/{len(files)}] {p.name} "
                          f"+{count} rows")
            # CMP-AUD-027: a valid-header file that contributes ZERO data rows
            # adds no [route, …] row, so the route would silently VANISH from
            # coverage — the comparison could then certify a clean match while a
            # whole route present on one side is invisible. The statewide census
            # (756 real per-route exports across the four flat families; min 1
            # data row) found NO header-only file, so a data-less per-route
            # export is anomalous. Disclose it LOUDLY as an incomplete input,
            # naming the route, so its identity is never discarded (the sibling
            # of the CMP-AUD-030/031 skips). A route already in `seen_routes`
            # keeps the export honest if a real file for it follows.
            if count == 0:
                events.on_log(f"  [{label}] {p.name}: valid header but no data "
                              f"rows (route {route}); flagging incomplete")
                skipped.append(f"{label} {p.name}: route {route} has a valid "
                               "header but no data rows (the export may be "
                               "truncated)")
        finally:
            wb.close()
    if not rows:
        raise ValueError(
            f"No readable {report_name} files were found for the {label} "
            f"side in:\n{in_dir}")
    if skipped:
        events.on_log(f"  [{label}] note: {len(skipped)} file(s) skipped "
                      "(details above).")
    return rows, header, skipped


# The Ramp Summary's comparison columns: the consolidator's own field order
# and display labels (Source + the formula-only Audit columns excluded — the
# comparison recomputes everything itself).
_RS_FIELDS = [(col, disp) for group, cols in _rs.GROUPS
              for col, disp in cols
              if group not in ("Source", "Audit")]
RS_HEADER = ["Route"] + [disp for _col, disp in _RS_FIELDS]


def _load_ramp_summary_side(folder, label, events):
    """Parse every per-route Ramp Summary PDF on one side into per-route-shape
    rows ([route, *numeric fields] — the route IS the row key). Slow-ish
    (~1-2 s per PDF), so progress is logged per file and cancel is honored.
    Returns (rows, skipped) — `skipped` lists the PDFs that wouldn't parse, so
    the caller can flag the comparison as incomplete rather than silently
    dropping a route."""
    in_dir, files = _find_input_dir(folder, "ramp_summary", "*.pdf")
    if not files:
        raise ValueError(
            f"No Ramp Summary PDFs were found for the {label} side:\n{in_dir}\n\n"
            "Export the Ramp Summary report on that environment first.")
    rows = []
    skipped = []
    for i, p in enumerate(files, 1):
        if events.is_cancelled():
            raise ValueError("Cancelled by user.")
        try:
            record = _rs.parse_pdf(p)
        except Exception as e:
            events.on_log(f"  [{label}] {p.name}: could not parse "
                          f"({type(e).__name__}); skipping")
            log.warning("env compare: %s parse failed", p, exc_info=True)
            skipped.append(f"{label} {p.name}: could not parse "
                           f"({type(e).__name__})")
            continue
        if not _rs.record_has_data(record):
            # One-page / truncated PDF: a route but no ramp figures. Skip it
            # (an all-blank route would otherwise compare as a phantom match).
            events.on_log(f"  [{label}] {p.name}: no ramp data "
                          "(one-page / truncated PDF?); skipping")
            skipped.append(f"{label} {p.name}: no ramp data "
                           "(one-page / truncated PDF?)")
            continue
        route = _norm_route_key(record.get("route") or _route_from_name(p))
        rows.append([route] + [record.get(col) for col, _disp in _RS_FIELDS])
        events.on_log(f"  [{label}] [{i:>3}/{len(files)}] {p.name} "
                      f"(route {route})")
    if not rows:
        raise ValueError(
            f"No readable Ramp Summary PDFs were found for the {label} side "
            f"in:\n{in_dir}")
    if skipped:
        events.on_log(f"  [{label}] note: {len(skipped)} PDF(s) skipped "
                      "(details above).")
    return rows, skipped


# Intersection Summary's comparison columns: Total + one column per canonical
# category (the consolidator's category key as the header), in spec order.
_IS_FIELDS = [(c.slug, c.key) for c in _is._CATS]
IS_HEADER = ["Route", "Total Intersections"] + [key for _slug, key in _IS_FIELDS]


def _load_intersection_summary_side(folder, label, events):
    """Parse every per-route Intersection Summary XLSX on one side into per-route-
    shape rows ([route, total, *category counts] — the route IS the row key), reusing
    the consolidator's own block-walk parser (consolidate_intersection_summary
    .parse_route) so the two sides can't drift from each other or the consolidation.
    Returns (rows, skipped); raises ValueError when nothing is readable."""
    in_dir, files = _find_input_dir(folder, "intersection_summary", "*.xlsx")
    files = [p for p in files if not p.name.startswith("~$")]
    if not files:
        raise ValueError(
            f"No Intersection Summary files were found for the {label} side:\n{in_dir}"
            "\n\nExport the Intersection Summary report on that environment first.")
    rows, skipped = [], []
    for i, p in enumerate(files, 1):
        if events.is_cancelled():
            raise ValueError("Cancelled by user.")
        try:
            route, counts, total = _is.parse_route(str(p))
        except Exception as e:
            events.on_log(f"  [{label}] {p.name}: could not parse "
                          f"({type(e).__name__}); skipping")
            log.warning("env compare: %s parse failed", p, exc_info=True)
            skipped.append(f"{label} {p.name}: could not parse ({type(e).__name__})")
            continue
        if not _is.record_has_data({"counts": counts}):
            events.on_log(f"  [{label}] {p.name}: no intersection data; skipping")
            skipped.append(f"{label} {p.name}: no intersection data")
            continue
        rkey = _norm_route_key(route)
        rows.append([rkey, total] + [counts.get(slug, 0) for slug, _k in _IS_FIELDS])
        events.on_log(f"  [{label}] [{i:>3}/{len(files)}] {p.name} (route {rkey})")
    if not rows:
        raise ValueError(
            f"No readable Intersection Summary files were found for the {label} "
            f"side in:\n{in_dir}")
    if skipped:
        events.on_log(f"  [{label}] note: {len(skipped)} file(s) skipped (details above).")
    return rows, skipped


def _non_negative_result_count(result, field, report_name):
    """Read a producer-owned loss count without silently normalizing bad truth."""
    value = getattr(result, field, 0)
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise ValueError(
            f"The {report_name} PDF converter returned an invalid {field} count; "
            "the comparison was not created.")
    return value


def _pdf_loaded_side(result, loaded, *, label, report_name, source_pdf_count):
    """Merge one PDF converter outcome with its converted-XLSX loader outcome.

    The PDF consolidators can successfully publish a *partial* workbook.  Their
    completion and loss counters are producer truth and must not disappear merely
    because the converted XLSX files that remain are internally consistent.  Keep
    those claims structured; do not scrape the consolidator's display-oriented
    ``summary_lines``.
    """
    rows, header, converted_skips = loaded
    producer_completion = getattr(result, "completion", None) or "unknown"
    if producer_completion in ("failed", "no_data", "cancelled"):
        raise ValueError(
            f"The {report_name} PDF converter reported {producer_completion!r} "
            "completion, so that side cannot be compared.")
    producer_skipped = _non_negative_result_count(
        result, "skipped_inputs", report_name)
    producer_failed = _non_negative_result_count(
        result, "failed_inputs", report_name)
    converted_skips = tuple(str(item) for item in converted_skips)

    warnings = list(converted_skips)
    failures = []
    if producer_skipped:
        warnings.append(
            f"{label}: {report_name} PDF conversion reported "
            f"{producer_skipped} skipped input item(s).")
    if producer_failed:
        failures.append(
            f"{label}: {report_name} PDF conversion reported "
            f"{producer_failed} failed input file(s).")
    if (producer_completion != "complete"
            and not producer_skipped and not producer_failed):
        warnings.append(
            f"{label}: {report_name} PDF conversion completion was "
            f"{producer_completion!r}, not complete.")

    # Rows exist (the caller already rejected a failed/no-readable conversion), so
    # any producer loss or converted-file loss is a comparable-but-partial side.
    complete = (producer_completion == "complete"
                and producer_skipped == 0 and producer_failed == 0
                and not converted_skips)
    return LoadedSide(
        rows=tuple(tuple(row) for row in rows),
        declared_schema=tuple(header),
        route_universe=tuple(sorted({str(row[0]) for row in rows if row})),
        completion="complete" if complete else "partial",
        warnings=tuple(warnings),
        failures=tuple(failures),
        skipped_inputs=producer_skipped + len(converted_skips),
        failed_inputs=producer_failed,
        raw_identity_claims={
            "pdf_consolidation": {
                "status": str(getattr(result, "status", "") or ""),
                "completion": str(producer_completion),
                "skipped_inputs": producer_skipped,
                "failed_inputs": producer_failed,
            },
        },
        display_metrics={
            "source_pdf_count": source_pdf_count,
            "loaded_row_count": len(rows),
        },
    )


def _load_highway_log_pdf_side(folder, label, events):
    """Parse one side's Highway Log PDFs (folder/highway_log_pdf/*.pdf) into
    consolidated-shape 31-column rows: convert them to per-route XLSX with the HL-PDF
    consolidator's own parser in a temp dir, then read those flat like any XLSX side.
    Returns a LoadedSide. The PDF is the ACCURATE Highway Log source (the vendor
    Excel drops rows), so this is the preferred cross-env Highway Log."""
    import consolidate_tsmis_highway_log_pdf as _hlpdf
    import highway_log_columns as hlc
    in_dir, pdfs = _find_input_dir(folder, _hlpdf.SUBDIR, "*.pdf")
    if not pdfs:
        raise ValueError(
            f"No Highway Log (PDF) files were found for the {label} side:\n{in_dir}"
            "\n\nExport the Highway Log (PDF) report on that environment first.")
    conv = Path(tempfile.mkdtemp(prefix="hlpdf_env_conv_"))
    combined_dir = Path(tempfile.mkdtemp(prefix="hlpdf_env_out_"))
    try:
        res = _hlpdf.consolidate(events=events, confirm_overwrite=lambda _p: True,
                                 input_dir=in_dir, out_path=combined_dir / "_combined.xlsx",
                                 converted_dir=conv)
        if res.status == "cancelled":
            raise ValueError("Cancelled by user.")
        if res.status != "ok":
            raise ValueError(res.message or "Could not parse the Highway Log PDFs.")
        # The per-route XLSX now sit in `conv` (the combined file is in combined_dir,
        # excluded). Read them flat with the corrected 31-column header pinned and
        # the SAME Highway Log projection the dedicated comparator uses
        # (CMP-AUD-047 — projection parity across every Highway Log entry point).
        loaded = _load_xlsx_side(
            conv, label, "_perroute_", _hlpdf.SHEET_NAME,
            "Highway Log (PDF)", events, expected_header=hlc.HEADER,
            value_normalizer=_hl._hl_normalize)
        return _pdf_loaded_side(
            res, loaded, label=label, report_name="Highway Log (PDF)",
            source_pdf_count=len(pdfs))
    finally:
        shutil.rmtree(conv, ignore_errors=True)
        shutil.rmtree(combined_dir, ignore_errors=True)


def _load_highway_detail_pdf_side(folder, label, events):
    """Parse one side's Highway Detail PDFs (folder/highway_detail_pdf/*.pdf)
    into consolidated-shape 34-column rows: convert them to per-route XLSX with
    the HD-PDF consolidator's own parser in a temp dir, then read those flat like
    any XLSX side. Returns a LoadedSide. The exact parallel of
    _load_intersection_detail_pdf_side below."""
    import consolidate_tsmis_highway_detail_pdf as _hdpdf
    import highway_detail_columns as hdc
    in_dir, pdfs = _find_input_dir(folder, _hdpdf.SUBDIR, "*.pdf")
    if not pdfs:
        raise ValueError(
            f"No Highway Detail (PDF) files were found for the {label} side:\n{in_dir}"
            "\n\nExport the Highway Detail (PDF) report on that environment first.")
    conv = Path(tempfile.mkdtemp(prefix="hdpdf_env_conv_"))
    combined_dir = Path(tempfile.mkdtemp(prefix="hdpdf_env_out_"))
    try:
        res = _hdpdf.consolidate(events=events, confirm_overwrite=lambda _p: True,
                                 input_dir=in_dir, out_path=combined_dir / "_combined.xlsx",
                                 converted_dir=conv)
        if res.status == "cancelled":
            raise ValueError("Cancelled by user.")
        if res.status != "ok":
            raise ValueError(res.message or "Could not parse the Highway Detail PDFs.")
        # The per-route XLSX now sit in `conv` (the combined file is in combined_dir,
        # excluded). Read them flat with the 34-column Highway Detail header pinned.
        loaded = _load_xlsx_side(
            conv, label, "_perroute_", _hdpdf.SHEET_NAME,
            "Highway Detail (PDF)", events, expected_header=hdc.HEADER)
        return _pdf_loaded_side(
            res, loaded, label=label, report_name="Highway Detail (PDF)",
            source_pdf_count=len(pdfs))
    finally:
        shutil.rmtree(conv, ignore_errors=True)
        shutil.rmtree(combined_dir, ignore_errors=True)


def _load_intersection_detail_pdf_side(folder, label, events):
    """Parse one side's Intersection Detail PDFs (folder/intersection_detail_pdf/*.pdf)
    into consolidated-shape 36-column rows: convert them to per-route XLSX with the
    Int-Detail-PDF consolidator's own parser in a temp dir, then read those flat like
    any XLSX side. Returns a LoadedSide. The exact parallel of
    _load_highway_log_pdf_side — both PDF sides are parsed offline the same way the
    PDF-vs-TSN comparison parses them, so the cross-env comparison is consistent."""
    import consolidate_tsmis_intersection_detail_pdf as _idpdf
    import intersection_detail_columns as idc
    in_dir, pdfs = _find_input_dir(folder, _idpdf.SUBDIR, "*.pdf")
    if not pdfs:
        raise ValueError(
            f"No Intersection Detail (PDF) files were found for the {label} side:\n{in_dir}"
            "\n\nExport the Intersection Detail (PDF) report on that environment first.")
    conv = Path(tempfile.mkdtemp(prefix="idpdf_env_conv_"))
    combined_dir = Path(tempfile.mkdtemp(prefix="idpdf_env_out_"))
    try:
        res = _idpdf.consolidate(events=events, confirm_overwrite=lambda _p: True,
                                 input_dir=in_dir, out_path=combined_dir / "_combined.xlsx",
                                 converted_dir=conv)
        if res.status == "cancelled":
            raise ValueError("Cancelled by user.")
        if res.status != "ok":
            raise ValueError(res.message or "Could not parse the Intersection Detail PDFs.")
        # The per-route XLSX now sit in `conv` (the combined file is in combined_dir,
        # excluded). Read them flat with the 36-column Intersection Detail header pinned.
        loaded = _load_xlsx_side(
            conv, label, "_perroute_", _idpdf.SHEET_NAME,
            "Intersection Detail (PDF)", events, expected_header=idc.HEADER)
        return _pdf_loaded_side(
            res, loaded, label=label, report_name="Intersection Detail (PDF)",
            source_pdf_count=len(pdfs))
    finally:
        shutil.rmtree(conv, ignore_errors=True)
        shutil.rmtree(combined_dir, ignore_errors=True)


def _load_highway_sequence_pdf_side(folder, label, events):
    """Parse one side's Highway Sequence PDFs (folder/highway_sequence_pdf/*.pdf)
    into consolidated-shape 9-column rows: convert them to per-route XLSX with the
    HSL-PDF consolidator's own parser in a temp dir, then read those flat like any
    XLSX side. Returns a LoadedSide. The exact parallel of
    _load_highway_detail_pdf_side; no expected_header pin — the converted files
    carry the Excel export's own header (with its two unnamed columns), exactly
    like the Excel side the HIGHWAY_SEQUENCE row reads."""
    import consolidate_tsmis_highway_sequence_pdf as _hslpdf
    in_dir, pdfs = _find_input_dir(folder, _hslpdf.SUBDIR, "*.pdf")
    if not pdfs:
        raise ValueError(
            f"No Highway Sequence (PDF) files were found for the {label} side:\n{in_dir}"
            "\n\nExport the Highway Sequence Listing (PDF) report on that environment first.")
    conv = Path(tempfile.mkdtemp(prefix="hslpdf_env_conv_"))
    combined_dir = Path(tempfile.mkdtemp(prefix="hslpdf_env_out_"))
    try:
        res = _hslpdf.consolidate(events=events, confirm_overwrite=lambda _p: True,
                                  input_dir=in_dir, out_path=combined_dir / "_combined.xlsx",
                                  converted_dir=conv)
        if res.status == "cancelled":
            raise ValueError("Cancelled by user.")
        if res.status != "ok":
            raise ValueError(res.message or "Could not parse the Highway Sequence PDFs.")
        # The per-route XLSX now sit in `conv` (the combined file is in combined_dir,
        # excluded). Read them flat like the Excel per-route files.
        loaded = _load_xlsx_side(
            conv, label, "_perroute_", _hslpdf.SHEET_NAME,
            "Highway Sequence (PDF)", events)
        return _pdf_loaded_side(
            res, loaded, label=label, report_name="Highway Sequence (PDF)",
            source_pdf_count=len(pdfs))
    finally:
        shutil.rmtree(conv, ignore_errors=True)
        shutil.rmtree(combined_dir, ignore_errors=True)


def _load_ramp_detail_pdf_side(folder, label, events):
    """Parse one side's Ramp Detail PDFs (folder/ramp_detail_pdf/*.pdf) into
    consolidated-shape rows: convert them to per-route XLSX with the RD-PDF
    consolidator's own parser in a temp dir, then read those flat like any XLSX
    side. Returns a LoadedSide. The exact parallel of
    _load_highway_sequence_pdf_side above; no expected_header pin — the
    converted files carry the Excel export's own (column-shifted) header plus
    the two print-only columns, identical on BOTH sides of a PDF-vs-PDF pair."""
    import consolidate_tsmis_ramp_detail_pdf as _rdpdf
    in_dir, pdfs = _find_input_dir(folder, _rdpdf.SUBDIR, "*.pdf")
    if not pdfs:
        raise ValueError(
            f"No Ramp Detail (PDF) files were found for the {label} side:\n{in_dir}"
            "\n\nExport the TSAR: Ramp Detail (PDF) report on that environment first.")
    conv = Path(tempfile.mkdtemp(prefix="rdpdf_env_conv_"))
    combined_dir = Path(tempfile.mkdtemp(prefix="rdpdf_env_out_"))
    try:
        res = _rdpdf.consolidate(events=events, confirm_overwrite=lambda _p: True,
                                 input_dir=in_dir, out_path=combined_dir / "_combined.xlsx",
                                 converted_dir=conv)
        if res.status == "cancelled":
            raise ValueError("Cancelled by user.")
        if res.status != "ok":
            raise ValueError(res.message or "Could not parse the Ramp Detail PDFs.")
        # The per-route XLSX now sit in `conv` (the combined file is in combined_dir,
        # excluded). Read them flat like the Excel per-route files.
        loaded = _load_xlsx_side(
            conv, label, "_perroute_", _rdpdf.SHEET_NAME,
            "Ramp Detail (PDF)", events)
        return _pdf_loaded_side(
            res, loaded, label=label, report_name="Ramp Detail (PDF)",
            source_pdf_count=len(pdfs))
    finally:
        shutil.rmtree(conv, ignore_errors=True)
        shutil.rmtree(combined_dir, ignore_errors=True)


def _coerce_loaded_side(value, *, declared_schema=None):
    """Accept the typed loader contract and the two historical tuple shapes."""
    if isinstance(value, LoadedSide):
        if not value.declared_schema and declared_schema is not None:
            return replace(value, declared_schema=tuple(declared_schema))
        if not value.declared_schema:
            raise ValueError(
                "The typed input loader did not declare its column schema; "
                "the comparison was not created.")
        return value

    if declared_schema is None:
        rows, header, skipped = value
    else:
        rows, skipped = value
        header = declared_schema
    skipped = tuple(str(item) for item in skipped)
    return LoadedSide(
        rows=tuple(tuple(row) for row in rows),
        declared_schema=tuple(header),
        completion="partial" if skipped else "complete",
        warnings=skipped,
        skipped_inputs=len(skipped),
    )


def _side_coverage_items(side, label):
    """Return categorized diagnostics plus the fail-closed coverage decision."""
    warnings = tuple(str(item) for item in side.warnings)
    failures = tuple(str(item) for item in side.failures)
    incomplete = (side.completion != "complete" or bool(warnings or failures)
                  or side.skipped_inputs > 0 or side.failed_inputs > 0)
    if incomplete and not warnings and not failures:
        warnings = (f"{label}: loaded-side completion was {side.completion!r}, "
                    "not complete.",)
    return warnings, failures, incomplete


def _side_coverage_diagnostic(role, label, side):
    return {
        "kind": "loaded_side_coverage",
        "role": role,
        "label": label,
        "completion": side.completion,
        "skipped_inputs": side.skipped_inputs,
        "failed_inputs": side.failed_inputs,
        "raw_identity_claims": dict(side.raw_identity_claims),
        "display_metrics": dict(side.display_metrics),
    }


def _apply_pdf_coverage(result, sides):
    """Attach side coverage only when no artifact generation was published.

    Successful generations receive this truth inside run_compare, before commit.
    This fallback preserves diagnostics on pre-publication terminal failures.
    """
    total_skipped = sum(side.skipped_inputs for _role, _label, side in sides)
    total_failed = sum(side.failed_inputs for _role, _label, side in sides)
    warning_items = []
    failure_items = []
    incomplete = False
    diagnostics = []
    for role, label, side in sides:
        side_warnings, side_failures, side_incomplete = _side_coverage_items(
            side, label)
        warning_items.extend(side_warnings)
        failure_items.extend(side_failures)
        incomplete = incomplete or side_incomplete
        diagnostics.append(_side_coverage_diagnostic(role, label, side))

    # The legacy fields remain the compatibility surface, but they now carry the
    # exact aggregate counts instead of compare_core's diagnostic-string count.
    result.skipped_inputs = total_skipped
    result.failed_inputs = total_failed
    if getattr(result, "status", None) == "ok" and incomplete:
        result.completion = "partial"
        result.verdict = "diff"

    typed = getattr(result, "comparison_outcome", None)
    if isinstance(typed, ComparisonOutcome):
        # compare_core received these same items as coverage warnings.  Re-split
        # them here so failed inputs remain failures in the machine contract.
        typed_completion = ("partial" if incomplete and typed.status == "ok"
                            else typed.completion)
        typed_verdict = ("diff" if incomplete and typed.status == "ok"
                         else typed.verdict)
        retained_warnings = [item for item in typed.warnings
                             if item not in failure_items]
        retained_failures = list(typed.failures)
        for item in warning_items:
            if item not in retained_warnings:
                retained_warnings.append(item)
        for item in failure_items:
            if item not in retained_failures:
                retained_failures.append(item)
        result.comparison_outcome = replace(
            typed,
            completion=typed_completion,
            verdict=typed_verdict,
            warnings=tuple(retained_warnings),
            failures=tuple(retained_failures),
            coverage_diagnostics=(
                tuple(typed.coverage_diagnostics) + tuple(diagnostics)),
        )
    else:
        # The artifact transaction can fail after both sides were loaded but
        # before compare_core's typed result is returned (validation, late alias,
        # destination ownership).  Preserve the known side truth even then.
        status = getattr(result, "status", None)
        if status == "cancelled":
            completion = "cancelled"
        elif status == "error":
            completion = "failed"
        elif status == "ok":
            # The coarse status is trustworthy, but no typed engine counts exist.
            # Keep completion/verdict unknown rather than manufacturing success.
            completion = "unknown"
        else:
            status, completion = "unknown", "unknown"
        if status in ("error", "cancelled"):
            # Keep the legacy and typed terminal axes consistent.
            result.completion = completion
        elif status == "ok":
            # No engine-owned typed counts exist, so neither compatibility field
            # may retain or manufacture a successful comparison claim.
            result.completion = None
            result.verdict = None
        terminal_failures = list(failure_items)
        message = str(getattr(result, "message", "") or "")
        if status == "error" and message and message not in terminal_failures:
            terminal_failures.insert(0, message)
        result.comparison_outcome = ComparisonOutcome(
            status=status,
            completion=completion,
            verdict="unknown",
            counts=ComparisonCounts(),
            warnings=tuple(warning_items),
            failures=tuple(terminal_failures),
            coverage_diagnostics=tuple(diagnostics),
        )
    return result


# Source Files provenance (cross-env / baseline): the per-route export prefix + ext
# per report subdir. Both sides are per-route TSMIS exports and the loader prepends
# the route, so each row's source filename is reconstructed from column 0.
_SOURCE_FILE_SPEC = {
    "ramp_detail": ("tsar_ramp_detail", "xlsx"),
    "highway_sequence": ("highway_sequence", "xlsx"),
    "highway_detail": ("highway_detail", "xlsx"),
    "highway_log": ("highway_log", "xlsx"),
    "intersection_detail": ("intersection_detail", "xlsx"),
    "ramp_detail_pdf": ("tsar_ramp_detail", "pdf"),
    "highway_sequence_pdf": ("highway_sequence", "pdf"),
    "highway_detail_pdf": ("highway_detail", "pdf"),
    "intersection_detail_pdf": ("intersection_detail", "pdf"),
    "highway_log_pdf": ("highway_log", "pdf"),
}


def _compose_source_files(sc, subdir, rows_a, rows_b, la, lb):
    """Compose the default 'Source Files' sheet onto `sc` for a cross-env comparison
    (both sides are per-route TSMIS exports). No-op for subdirs without a known spec
    (e.g. the Ramp/Intersection Summary per-route parsers)."""
    spec = _SOURCE_FILE_SPEC.get(subdir)
    if spec is None:
        return sc
    prefix, ext = spec
    prev = sc.extra_sheet_writer

    def _writer(wb, ctx):
        if prev is not None:
            prev(wb, ctx)
        ctc.write_source_files_sheet(wb, [
            (la, rows_a, ctc.source_files_from_rows(rows_a, prefix, ext)),
            (lb, rows_b, ctc.source_files_from_rows(rows_b, prefix, ext)),
        ])

    return replace(sc, extra_sheet_writer=_writer)


# ---------------------------------------------------------------------------
# Per-report adapters
# ---------------------------------------------------------------------------

class EnvCompare:
    """One report type's cross-environment comparison (a COMPARE_REPORTS
    "folders" entry): compare_folders(dir_a, dir_b, out_path, …) +
    suggest_name(dir_a, dir_b)."""

    def __init__(self, key, report_name, subdir, sheet_name=None,
                 expected_header=None, base_schema=None, key_col=None,
                 force_header=None, side_loader=None, agg_header=None,
                 flat_pdf_loader=None, physical_key_builder=None,
                 value_normalizer=None, header_canonicalizer=None):
        self.key = key                        # "ramp_summary" | "ramp_detail" | …
        self.REPORT_NAME = report_name
        self.subdir = subdir
        self.sheet_name = sheet_name          # None = an aggregate (PDF/category) path
        self.expected_header = expected_header
        # Aggregate-per-route path: a custom (folder,label,events)->(rows,skipped)
        # loader that yields ONE [route, *fields] row per route (route is the key,
        # has_route=False), with `agg_header` as the schema header. Used by the
        # category-summary reports (Intersection Summary) whose per-route sheet
        # isn't a flat header+rows table. (Ramp Summary keeps its own PDF path.)
        self.side_loader = side_loader
        self.agg_header = agg_header
        # Flat path (has_route=True) whose source is PDFs, not XLSX: a
        # (folder,label,events)->(rows,header,skipped) loader used instead of
        # _load_xlsx_side. Highway Log (PDF) uses it to parse each side's PDFs.
        self.flat_pdf_loader = flat_pdf_loader
        self.base_schema = base_schema        # report-specific schema extras
        # When set, the DISPLAY header is forced to this (the loaded files are
        # accepted as-is and compared by POSITION). Highway Log uses it so the
        # vendor-mislabeled Excel exports still compare but show CORRECTED labels.
        self.force_header = force_header
        # The column NAME to key rows on, when the report's FIRST column is too
        # coarse to identify a row (e.g. Highway Sequence / Ramp Detail lead
        # with County / a district-county-route Location that repeats for
        # hundreds of rows; the postmile "PM" column is the real identity).
        # Resolved to a header index per loaded layout; falls back to the first
        # column (the original behavior) when the named column isn't present.
        self.key_col = key_col
        # CMP-AUD-045 (opt-in): callable(header, key_field) -> a compare_core
        # key_normalizer that builds the report's county-aware PhysicalKey from
        # each row, or None to degrade to the plain key column (logged). Wired
        # into every per-run schema so cross-environment pairing can never
        # match physically different locations that share a postmile.
        self.physical_key_builder = physical_key_builder
        # CMP-AUD-047 (opt-in): the report's OWN cell projection, applied by
        # the XLSX side loader instead of the generic normalize_value so the
        # cross-environment verdicts match the dedicated comparator's
        # (Highway Log collapses the export's tab/newline padding).
        self.value_normalizer = value_normalizer
        # CMP-AUD-048 (opt-in): callable(header) -> the CANONICAL header for a
        # recognized layout edition, or None for an unrecognized layout. When
        # set, both sides canonicalize independently BEFORE layout equality —
        # the documented canonical/vendor editions compare and display the
        # corrected names, while an unrecognized same-width header is refused
        # instead of compared positionally on faith.
        self.header_canonicalizer = header_canonicalizer

    def suggest_name(self, dir_a, dir_b):
        la, lb = _side_labels(Path(dir_a), Path(dir_b))
        safe = lambda s: re.sub(r"[^\w\-]+", "_", s).strip("_")
        # The generated-on date (A1): the side labels already carry each side's
        # export date + src-env, so this stamps WHEN the comparison was built.
        return (f"{safe(la)}_vs_{safe(lb)}_"
                f"{safe(self.REPORT_NAME)}_Comparison {today_str()}.xlsx")

    def _resolve_key_field(self, header):
        """Index of the configured key column in this loaded header, matched on
        the stripped, case-folded name. No configured key (`key_col=None`) means
        the first column is the legitimate identity (flat route-keyed reports).

        CMP-AUD-028: a CONFIGURED key column that is ABSENT is fail-closed. It
        used to log and fall back to column 0, so two malformed key-less
        workbooks paired on their first column and returned a clean MATCH. The
        raise is normalized to a typed error at the single `_schema` call site."""
        if not self.key_col:
            return 0
        want = self.key_col.strip().casefold()
        for i, name in enumerate(header):
            if name is not None and str(name).strip().casefold() == want:
                return i
        raise ValueError(
            f"The {self.REPORT_NAME} files have no '{self.key_col}' column — "
            "that is the row identity this comparison pairs on, so the two "
            "sides cannot be reliably matched. Re-export them from a supported "
            "app version.")

    def _schema(self, header, la, lb):
        # Force the corrected display header when configured (Highway Log); the
        # data is positional, so relabeling for display can't shift columns.
        if self.force_header is not None and len(self.force_header) == len(header):
            header = list(self.force_header)
        base = self.base_schema or CompareSchema(
            report_name=self.REPORT_NAME, header=header,
            id_noun="row", id_noun_plural="rows")
        widths = dict(base.data_widths) or {header[0]: 12}
        if "Description" in header and "Description" not in widths:
            widths["Description"] = 26
        cmp_widths = dict(base.cmp_widths)
        if "Description" in header and "Description" not in cmp_widths:
            cmp_widths["Description"] = 30
        key_field = self._resolve_key_field(header)
        key_normalizer = base.key_normalizer
        if self.physical_key_builder is not None:
            key_normalizer = self.physical_key_builder(header, key_field)
        return replace(base, header=header, side_a=la, side_b=lb,
                       sides_noun="environments",
                       data_widths=widths, cmp_widths=cmp_widths,
                       key_field=key_field, key_normalizer=key_normalizer,
                       one_sided_note_extra="", trim_note_extra="")

    def _effective_source_files(self, folder):
        """Files this adapter can actually load from one selected side.

        Capturing only the selected root misses an external hardlink to a
        report file inside it.  Mirror each loader's narrow discovery pattern
        so the alias guard sees the real input objects without stat-walking
        unrelated report trees in a Matrix environment folder.
        """
        use_pdf = (self.flat_pdf_loader is not None
                   or (self.sheet_name is None and self.side_loader is None))
        _input_dir, files = _find_input_dir(
            folder, self.subdir, "*.pdf" if use_pdf else "*.xlsx")
        return tuple(files)

    @comparison_result_boundary
    def compare_folders(self, dir_a, dir_b, out_path, events=None,
                        confirm_overwrite=None, mode="formulas", labels=None,
                        commit_guard=None):
        """Compare the report's per-route files in run folder `dir_a` against
        `dir_b` and write the discrepancy workbook(s) to `out_path`. Returns
        a ConsolidateResult (same contract as the consolidators). `labels`
        overrides the two derived side names (still capped + kept distinct) —
        the baseline matrix passes explicit ones because the Everything store's
        folder shape ("All Reports (current)/<src-env>") derives a side label
        confusingly close to the run-folder one."""
        events = events or Events()
        if not _XLSX_OK:
            return ConsolidateResult(
                status="error",
                message="Required components are missing (openpyxl).")
        if (self.sheet_name is None and self.side_loader is None
                and not getattr(_rs, "_DEPS_OK", False)):
            return ConsolidateResult(
                status="error",
                message="Required components are missing (pdfplumber).")
        dir_a, dir_b = Path(dir_a), Path(dir_b)
        for side, d in (("first", dir_a), ("second", dir_b)):
            if not d.is_dir():
                return ConsolidateResult(
                    status="error",
                    message=f"The {side} folder doesn't exist:\n{d}")
        if dir_a.resolve() == dir_b.resolve():
            return ConsolidateResult(
                status="error",
                message="Pick two DIFFERENT folders — both sides point at "
                        f"the same one:\n{dir_a}")
        destinations = artifact_store.comparison_output_paths(out_path, mode)
        files_a = self._effective_source_files(dir_a)
        files_b = self._effective_source_files(dir_b)
        discovered_paths = (*files_a, *files_b)
        discovered_set = artifact_store.canonical_path_identities(discovered_paths)
        source_paths = (dir_a, dir_b, *discovered_paths)
        try:
            captured_sources = artifact_store.capture_source_identities(
                source_paths)
            # CMP-AUD-076: the exact discovered member census per side, statted
            # BEFORE any loader reads (folder-kind provenance: the census IS the
            # effective input identity; a per-member content digest would be a
            # full re-read of hundreds of files — the discovery-set tripwire +
            # captured identities guard the read window instead).
            prov_sides = [
                {"kind": "folder",
                 "selection": str(d.resolve(strict=False)),
                 "member_count": len(files),
                 "members": _member_census(files)}
                for d, files in ((dir_a, files_a), (dir_b, files_b))]
        except ValueError as e:
            return ConsolidateResult(status="error", message=str(e))

        def alias_error():
            current_discovered = (*self._effective_source_files(dir_a),
                                  *self._effective_source_files(dir_b))
            if (artifact_store.canonical_path_identities(current_discovered)
                    != discovered_set):
                return ConsolidateResult(
                    status="error",
                    message=("Refusing to write the comparison: the discovered "
                             "source file set changed while inputs were loading. "
                             "Re-run after the source folders are stable."))
            try:
                artifact_store.ensure_outputs_do_not_alias_sources(
                    destinations, source_paths,
                    captured_sources=captured_sources,
                    require_sources_current=True)
            except ValueError as e:
                return ConsolidateResult(status="error", message=str(e))
            return None

        blocked = alias_error()
        if blocked is not None:
            return blocked

        if labels is not None:
            la, lb = _cap_label(str(labels[0])), _cap_label(str(labels[1]))
            if la == lb:                     # sheet names must differ
                la, lb = _cap_label(f"{la} (A)"), _cap_label(f"{lb} (B)")
        else:
            la, lb = _side_labels(dir_a, dir_b)

        events.on_log("=" * 60)
        events.on_log(f"{self.REPORT_NAME} Comparison — {la} vs {lb}")
        events.on_log("=" * 60)
        events.on_log(f"{la}: {dir_a}")
        events.on_log(f"{lb}: {dir_b}")
        events.on_log("")

        try:
            if self.side_loader is not None:  # aggregate per-route XLSX (Intersection Summary)
                loaded_a = _coerce_loaded_side(
                    self.side_loader(dir_a, la, events),
                    declared_schema=self.agg_header)
                loaded_b = _coerce_loaded_side(
                    self.side_loader(dir_b, lb, events),
                    declared_schema=self.agg_header)
                header, has_route = list(self.agg_header), False
            elif self.sheet_name is None:     # Ramp Summary: PDFs, route-keyed
                loaded_a = _coerce_loaded_side(
                    _load_ramp_summary_side(dir_a, la, events),
                    declared_schema=RS_HEADER)
                loaded_b = _coerce_loaded_side(
                    _load_ramp_summary_side(dir_b, lb, events),
                    declared_schema=RS_HEADER)
                header, has_route = list(RS_HEADER), False
            else:                             # flat (has_route=True): XLSX, or PDF-sourced
                def _flat_load(d, lab):
                    if self.flat_pdf_loader is not None:
                        return self.flat_pdf_loader(d, lab, events)
                    return _load_xlsx_side(d, lab, self.subdir, self.sheet_name,
                                           self.REPORT_NAME, events,
                                           expected_header=self.expected_header,
                                           value_normalizer=self.value_normalizer)
                loaded_a = _coerce_loaded_side(_flat_load(dir_a, la))
                loaded_b = _coerce_loaded_side(_flat_load(dir_b, lb))
                header = list(loaded_a.declared_schema)
                header_b = list(loaded_b.declared_schema)
                if self.header_canonicalizer is not None:
                    # CMP-AUD-048: each side canonicalizes INDEPENDENTLY, so
                    # one canonical-labelled and one vendor-labelled export of
                    # the same layout compare (displaying corrected names),
                    # while an unrecognized header is refused by NAME rather
                    # than trusted positionally.
                    canon_a = self.header_canonicalizer(header)
                    canon_b = self.header_canonicalizer(header_b)
                    if canon_a is None or canon_b is None:
                        bad = la if canon_a is None else lb
                        return ConsolidateResult(
                            status="error",
                            message=(f"The {bad} side's {self.REPORT_NAME} "
                                     "files do not use a recognized "
                                     f"{self.REPORT_NAME} column layout — "
                                     "re-export them with a supported app "
                                     "version."))
                    header, header_b = list(canon_a), list(canon_b)
                if header != header_b:
                    return ConsolidateResult(
                        status="error",
                        message=(f"The two folders' {self.REPORT_NAME} files "
                                 "have different column layouts — compare "
                                 "exports made by the same app version."))
                # rows are consolidated-shape ([route, *row]); the schema
                # header stays the per-route column list (the engine adds
                # the Route column itself in the has_route layout).
                has_route = True
        except ValueError as e:
            msg = str(e)
            status = "cancelled" if msg == "Cancelled by user." else "error"
            return ConsolidateResult(status=status, message=msg)

        # Files unreadable on EITHER side make the comparison incomplete — pass
        # them through so the verdict can't certify a clean match and the
        # workbook + summary list exactly what was left out.
        rows_a, rows_b = loaded_a.rows, loaded_b.rows
        warn_a, fail_a, incomplete_a = _side_coverage_items(loaded_a, la)
        warn_b, fail_b, incomplete_b = _side_coverage_items(loaded_b, lb)
        warnings = list(warn_a + warn_b)
        failures = list(fail_a + fail_b)
        total_skipped = loaded_a.skipped_inputs + loaded_b.skipped_inputs
        total_failed = loaded_a.failed_inputs + loaded_b.failed_inputs
        input_completion = ("partial" if incomplete_a or incomplete_b
                            else "complete")
        coverage_diagnostics = (
            _side_coverage_diagnostic("side_a", la, loaded_a),
            _side_coverage_diagnostic("side_b", lb, loaded_b),
        )
        if warnings or failures:
            events.on_log(
                f"⚠ Incomplete input coverage across both sides "
                f"({total_skipped} skipped, {total_failed} failed) — "
                "the comparison will be flagged incomplete.")

        events.on_log("")
        # A folder comparison may spend minutes loading a statewide side. Check
        # again immediately before the direct run_compare write so a late alias
        # cannot bypass this lower boundary.
        blocked = alias_error()
        if blocked is not None:
            return blocked
        try:
            sc = self._schema(header, la, lb)
        except ValueError as e:
            # CMP-AUD-028: a configured identity column is mandatory. _schema
            # (via _resolve_key_field) refuses a header that lacks it rather
            # than silently keying on column 0 and certifying a false match.
            return ConsolidateResult(status="error", message=str(e))
        # Default provenance: both sides are per-route TSMIS exports, so name the
        # source file each row came from (from the route prepended at column 0).
        sc = _compose_source_files(sc, self.subdir, rows_a, rows_b, la, lb)
        prov_sides[0]["role"], prov_sides[1]["role"] = la, lb
        prov_display = {"recipe": {"report": self.REPORT_NAME,
                                   "banner": f"{self.REPORT_NAME} Comparison "
                                             f"— {la} vs {lb}"},
                        "inputs": prov_sides}
        result = artifact_store.commit_workbook(
            out_path,
            lambda tmp: run_compare(
                sc, rows_a, rows_b, has_route, tmp,
                events=events, confirm_overwrite=lambda _p: True,
                mode=mode, name_a=str(dir_a.name), name_b=str(dir_b.name),
                warnings=warnings, commit_guard=commit_guard,
                 input_completion=input_completion,
                 skipped_inputs=total_skipped, failed_inputs=total_failed,
                 failures=failures,
                 provenance=prov_display,
                 coverage_diagnostics=coverage_diagnostics),
            twin=(mode == "both"), expect_sheet="Comparison",
            confirm_overwrite=confirm_overwrite,
            source_paths=source_paths,
            captured_sources=captured_sources,
            commit_guard=commit_guard,
            requested_mode=mode)
        # CMP-AUD-076: bind the pre-read member census to the committed
        # generation, beside the workbook (additive evidence; no-op unless a
        # generation actually committed).
        ctc.write_comparison_provenance(
            result, out_path, report=self.REPORT_NAME,
            banner=f"{self.REPORT_NAME} Comparison — {la} vs {lb}",
            inputs=prov_sides, commit_guard=commit_guard)
        if (self.flat_pdf_loader is not None
                and getattr(result, "artifact_generation", None) is None):
            return _apply_pdf_coverage(
                result, (("side_a", la, loaded_a), ("side_b", lb, loaded_b)))
        return result


# Highway Log: keep the Med Wid rule, the sample's widths, and the column
# tooltips + Legend (inherited from _hl._SCHEMA); the other XLSX reports lock
# their layout from the files. The vendor Excel exports carry the OLD mislabeled
# header, so the cross-env loader accepts the files as-is (lock from files) and
# force_header relabels the display to the CORRECTED header (positional).
# key_normalizer is cleared: cross-env compares two TSMIS exports (SAME roadbed
# encoding — both suffix the Location), so the roadbed-unifying key is unnecessary
# and would only perturb the validated cross-env output. It is a TSMIS-vs-TSN tool.
_HL_BASE = replace(_hl._SCHEMA, one_sided_note_extra="", trim_note_extra="",
                   key_normalizer=None)

RAMP_SUMMARY = EnvCompare(
    "ramp_summary", "Ramp Summary", "ramp_summary",
    base_schema=CompareSchema(
        report_name="Ramp Summary", header=RS_HEADER,
        id_noun="route", id_noun_plural="routes",
        scope_flat="All routes (one row per route)"))
def _ramp_detail_env_keys(header, key_field):
    """CMP-AUD-045: the Ramp Detail cross-env key builder — the same D4
    county-aware PhysicalKey the vs-TSN paths bake into their rows, built here
    per row from the export's own columns (County from the Location column;
    the cells before the PM column and after the Date column are the PM prefix
    and suffix, conserved as raw claims only). Degrades to the plain key column
    (logged) when the layout doesn't expose Location or a usable PM position —
    layout drift falls back to the old behavior rather than crashing, exactly
    like _resolve_key_field."""
    import compare_ramp_detail_tsn as _rd
    loc_field = next((i for i, name in enumerate(header)
                      if name is not None
                      and str(name).strip().casefold() == "location"), None)
    if loc_field is None or key_field == 0:
        log.warning("env compare ramp_detail: no Location column / key column "
                    "in header %r; falling back to the plain PM key", header)
        return None

    def normalizer(row, off, kf):
        route = "" if row[0] is None else str(row[0])
        loc = _rd._raw_text(row[off + loc_field]).strip()
        _district, county = _rd._dist_cnty(loc)
        pm_raw = row[off + kf]

        def cell(field):
            i = off + field
            return _rd._raw_text(row[i]) if 0 <= field < len(header) and i < len(row) else ""
        return _rd._physical_pm_key(route, county, pm_raw, (
            ("route", route), ("location", loc),
            ("postmile_prefix", cell(kf - 1)),
            ("postmile", _rd._raw_text(pm_raw)),
            ("postmile_suffix", cell(kf + 2))), f"Location {loc!r}")
    return normalizer


RAMP_DETAIL = EnvCompare(
    "ramp_detail", "Ramp Detail", "ramp_detail", sheet_name="TSAR - Ramp Detail",
    key_col="PM", physical_key_builder=_ramp_detail_env_keys)


def _hsl_unnamed_col(label):
    """True for a header cell the export left BLANK (compare_env labels those
    '(col X)'): the Highway Sequence prefix/suffix columns are exactly the two
    unnamed cells flanking PM in the real export layout."""
    return label is None or re.fullmatch(
        r"\(col [A-Z]+\)", str(label).strip()) is not None


def _highway_sequence_env_keys(header, key_field):
    """CMP-AUD-045: the Highway Sequence cross-env key builder — the same
    oracle-approved (Route, County, complete glued postmile) PhysicalKey the
    vs-TSN paths bake into their rows, built here per row from the export's own
    columns (County by name; the UNNAMED cells immediately before/after the PM
    column are the realignment prefix and equate suffix, glued INTO the
    canonical postmile — HSL's convention, unlike Ramp Detail's D4). Degrades
    to the plain key column (logged) when the layout doesn't expose County or
    the unnamed prefix/suffix flanks — layout drift falls back to the old
    behavior rather than gluing arbitrary neighbor columns into identity."""
    import compare_highway_sequence_tsn as _hsl
    county_field = next((i for i, name in enumerate(header)
                         if name is not None
                         and str(name).strip().casefold() == "county"), None)
    if county_field is None or key_field == 0:
        log.warning("env compare highway_sequence: no County column / key column "
                    "in header %r; falling back to the plain PM key", header)
        return None
    if (key_field - 1 <= 0 or key_field + 1 >= len(header)
            or not _hsl_unnamed_col(header[key_field - 1])
            or not _hsl_unnamed_col(header[key_field + 1])):
        log.warning("env compare highway_sequence: PM at %d is not flanked by "
                    "the unnamed prefix/suffix columns in header %r; falling "
                    "back to the plain PM key", key_field, header)
        return None

    def normalizer(row, off, kf):
        route = "" if row[0] is None else str(row[0])
        county_raw = row[off + county_field]

        def cell(field):
            i = off + field
            return _hsl._raw_text(row[i]) if 0 <= field < len(header) and i < len(row) else ""
        prefix, pm_raw, suffix = cell(kf - 1), row[off + kf], cell(kf + 1)
        return _hsl._physical_pm_key(
            route, county_raw,
            _hsl._glue_pm(prefix, pm_raw, suffix),
            (("route", route), ("county", _hsl._raw_text(county_raw)),
             ("postmile_prefix", prefix),
             ("postmile", _hsl._raw_text(pm_raw)),
             ("postmile_suffix", suffix)),
            f"the {row[0]!r} export row")
    return normalizer


HIGHWAY_SEQUENCE = EnvCompare(
    "highway_sequence", "Highway Sequence", "highway_sequence",
    sheet_name="Highway Locations", key_col="PM",
    physical_key_builder=_highway_sequence_env_keys)
def _hl_canonical_header(header):
    """Either supported Highway Log edition (canonical or vendor labels, with
    or without a leading Route) -> the canonical corrected header for layout
    equality; None = not a recognized Highway Log layout (CMP-AUD-048)."""
    import highway_log_columns as hlc
    has_route = hlc.recognize(list(header))
    if has_route is None:
        return None
    return ([hlc.ROUTE_COL] + list(hlc.HEADER)) if has_route \
        else list(hlc.HEADER)


HIGHWAY_LOG = EnvCompare(
    "highway_log", "Highway Log", "highway_log", sheet_name=_hl.SHEET_NAME,
    base_schema=_HL_BASE, force_header=_hl.EXPECTED_HEADER,
    value_normalizer=_hl._hl_normalize,
    header_canonicalizer=_hl_canonical_header)
# Intersection Summary: the per-route export is a CATEGORY-summary sheet (not a flat
# table), so it's compared the AGGREGATE way (one row of category counts per route,
# route-keyed) via the consolidator's own block-walk parser — the analog of Ramp
# Summary's PDF path, but XLSX.
INTERSECTION_SUMMARY = EnvCompare(
    "intersection_summary", "Intersection Summary", "intersection_summary",
    side_loader=_load_intersection_summary_side, agg_header=IS_HEADER,
    base_schema=CompareSchema(
        report_name="Intersection Summary", header=IS_HEADER,
        id_noun="route", id_noun_plural="routes",
        scope_flat="All routes (one row per route)"))
# Intersection Detail: a flat per-route XLSX (sheet "Intersection Detail"), keyed on
# the postmile, so it uses the standard flat path like Ramp Detail / Highway Sequence.
# (The export header is offset within each type/eff-date pair, but BOTH env sides share
# the identical layout, so the position-wise comparison is valid; intersections sharing
# a postmile pair by data similarity in compare_core. id_noun = intersection.)
def _intersection_detail_env_keys(header, key_field):
    """CMP-AUD-045: the Intersection Detail cross-env key builder — the same
    accepted ID-79 PhysicalKey the vs-TSN paths bake in, built per row from the
    export's own columns (County + base route from the Location column; the
    cell before the Post Mile column is the complete PP, PART of identity;
    the route suffix stays a conserved claim). Degrades to the plain key column
    (logged) on layout drift, like _resolve_key_field."""
    import compare_intersection_detail_tsn as _idt
    loc_field = next((i for i, name in enumerate(header)
                      if name is not None
                      and str(name).strip().casefold() == "location"), None)
    if loc_field is None or key_field == 0:
        log.warning("env compare intersection_detail: no Location column / key "
                    "column in header %r; falling back to the plain PM key",
                    header)
        return None

    def normalizer(row, off, kf):
        route = "" if row[0] is None else str(row[0])
        loc = row[off + loc_field]
        _district, county = _idt._dist_cnty(loc)
        _base, route_suffix = _idt._split_route(loc)
        pp = row[off + kf - 1]
        pm_raw = row[off + kf]
        def cell(field):
            i = off + field
            return _idt._raw_text(row[i]) if 0 <= field < len(header) and i < len(row) else ""
        return _idt._physical_id_key(route, county, pp, pm_raw, (
            ("route", route), ("route_suffix", route_suffix),
            ("location", _idt._raw_text(loc)),
            ("postmile_prefix", _idt._raw_text(pp)),
            ("postmile", _idt._raw_text(pm_raw)),
            ("postmile_suffix", cell(kf + 1))), f"Location {loc!r}")
    return normalizer


def _id_canonical_header(header):
    """Map either supported Intersection Detail site edition — the current
    (2026-07-17) labels or the 7.8/7.9 legacy labels, with or without a leading
    Route — to ONE canonical header, so a new-vs-old (pre/post July-2026)
    cross-env / baseline comparison aligns BY POSITION instead of refusing on the
    LABEL-ONLY change. The July-2026 update relabeled only ('P'->'PP', 'S'->'PS',
    the INT Type / INT Eff-Date labels realigned over their own values,
    'Ctrl T'->'Ctrl T Eff-Date', 'Xing P/S'->'Int PS'); every VALUE stayed in an
    identical column position (proven cell-for-cell), so position alignment is
    exact and any real data change (the Int St Eff-Date refresh, HG, the Location
    suffix) still surfaces as a genuine diff.

    Any OTHER header is returned UNCHANGED — the existing strict same-layout
    equality is preserved for every non-edition case (two identical custom
    headers still compare; two genuinely different layouts still refuse), so this
    only ADDS the cross-edition bridge and changes nothing else. Unlike Highway
    Log's `_hl_canonical_header` (CMP-AUD-048), it must not refuse an
    unrecognized layout: Intersection Detail carries no `force_header`, so its
    non-edition inputs are compared as-is."""
    import compare_intersection_detail_tsn as _idt
    h = [("" if c is None else str(c)).strip() for c in header]
    has_route = bool(h) and h[0] == "Route"
    body = h[1:] if has_route else h
    current, legacy = _idt._TSMIS_HEADER[1:], _idt._TSMIS_HEADER_LEGACY[1:]
    if body == list(current) or body == list(legacy):
        return (["Route"] + list(current)) if has_route else list(current)
    return list(header)


INTERSECTION_DETAIL = EnvCompare(
    "intersection_detail", "Intersection Detail", "intersection_detail",
    sheet_name="Intersection Detail", key_col="Post Mile",
    physical_key_builder=_intersection_detail_env_keys,
    header_canonicalizer=_id_canonical_header,
    base_schema=CompareSchema(
        report_name="Intersection Detail", header=["Post Mile"],
        id_noun="intersection", id_noun_plural="intersections", pair_noun="postmile"))
# Highway Log (PDF) cross-env: same Highway Log schema as the Excel row (Med Wid rule,
# corrected labels, ditto/roadbed), but BOTH sides are parsed from the app's own PDF
# export — the accurate Highway Log source (the vendor Excel drops rows), so this is
# the preferred cross-env Highway Log. flat_pdf_loader parses each side's PDFs first.
HIGHWAY_LOG_PDF = EnvCompare(
    "highway_log_pdf", "Highway Log (PDF)", "highway_log_pdf",
    sheet_name=_hl.SHEET_NAME, base_schema=_HL_BASE,
    force_header=_hl.EXPECTED_HEADER, flat_pdf_loader=_load_highway_log_pdf_side)
# Intersection Detail (PDF) cross-env: the same flat per-route shape as the Excel
# Intersection Detail row (sheet "Intersection Detail", keyed on Post Mile), but BOTH
# sides are parsed from the app's own PDF export — the exact parallel of HIGHWAY_LOG_PDF.
# No force_header: unlike the vendor-mislabeled Highway Log Excel, the Intersection
# Detail header is native, so the PDF and Excel sides already share the 36-col layout.
INTERSECTION_DETAIL_PDF = EnvCompare(
    "intersection_detail_pdf", "Intersection Detail (PDF)", "intersection_detail_pdf",
    sheet_name="Intersection Detail", key_col="Post Mile",
    base_schema=CompareSchema(
        report_name="Intersection Detail (PDF)", header=["Post Mile"],
        id_noun="intersection", id_noun_plural="intersections", pair_noun="postmile"),
    flat_pdf_loader=_load_intersection_detail_pdf_side)
# Highway Detail: a flat per-route XLSX (sheet "Highway Detail", 34 correct labels),
# keyed on the glued Post Mile — both env sides share the identical TSMIS encoding
# (prefix/roadbed/equation glued the same way), so the standard flat path applies and
# no roadbed canonicalization is needed (that is a TSMIS-vs-TSN tool).
HIGHWAY_DETAIL = EnvCompare(
    "highway_detail", "Highway Detail", "highway_detail",
    sheet_name="Highway Detail", key_col="Post Mile",
    base_schema=CompareSchema(
        report_name="Highway Detail", header=["Post Mile"],
        id_noun="location", id_noun_plural="locations", pair_noun="postmile"))
# Highway Detail (PDF) cross-env: the same flat shape, but BOTH sides are parsed from
# the app's own PDF export — the exact parallel of INTERSECTION_DETAIL_PDF.
HIGHWAY_DETAIL_PDF = EnvCompare(
    "highway_detail_pdf", "Highway Detail (PDF)", "highway_detail_pdf",
    sheet_name="Highway Detail", key_col="Post Mile",
    base_schema=CompareSchema(
        report_name="Highway Detail (PDF)", header=["Post Mile"],
        id_noun="location", id_noun_plural="locations", pair_noun="postmile"),
    flat_pdf_loader=_load_highway_detail_pdf_side)
# Highway Sequence (PDF) cross-env: the same flat per-route shape as the Excel
# HIGHWAY_SEQUENCE row (sheet "Highway Locations", keyed on PM), but BOTH sides
# are parsed from the app's own PDF export — the exact parallel of
# HIGHWAY_DETAIL_PDF above.
HIGHWAY_SEQUENCE_PDF = EnvCompare(
    "highway_sequence_pdf", "Highway Sequence (PDF)", "highway_sequence_pdf",
    sheet_name="Highway Locations", key_col="PM",
    flat_pdf_loader=_load_highway_sequence_pdf_side)
# Ramp Detail (PDF) cross-env: the same flat per-route shape as the Excel
# RAMP_DETAIL row (sheet "TSAR - Ramp Detail", keyed on PM) plus the two
# print-only columns (identical on both sides of a PDF-vs-PDF pair), but BOTH
# sides are parsed from the app's own PDF export — the exact parallel of
# HIGHWAY_SEQUENCE_PDF above.
RAMP_DETAIL_PDF = EnvCompare(
    "ramp_detail_pdf", "Ramp Detail (PDF)", "ramp_detail_pdf",
    sheet_name="TSAR - Ramp Detail", key_col="PM",
    flat_pdf_loader=_load_ramp_detail_pdf_side)

# Default save location for cross-environment comparison workbooks (the GUI
# aims its save dialog here; "Delete all reports" clears it).
DEFAULT_OUT_DIR = OUTPUT_ROOT / "comparisons"
