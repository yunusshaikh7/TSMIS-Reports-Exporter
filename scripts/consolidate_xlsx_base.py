"""Shared core for the XLSX-input consolidators.

Three reports export a single sheet with a header row (row 1) followed by data
rows (row 2+): TSAR: Ramp Detail, Highway Sequence Listing, and Highway Log.
Combining them is identical except for four values -- the input subfolder, the
output file name, the exact sheet name, and the friendly report name -- so that
common logic lives here once. A fix (or a tweak to the combined layout) now
benefits all three instead of drifting across near-duplicate copies.

The TSAR: Ramp Summary consolidator deliberately does NOT use this: it parses
PDFs with report-specific column/wrap logic, not XLSX, so it stays standalone.

Console-free, like the rest of the core: progress is reported via Events.on_log,
overwrite is confirmed through the confirm_overwrite(path)->bool callback, the run
honors events.is_cancelled(), and a ConsolidateResult is returned. Never
prints/prompts/exits. Third-party imports are guarded so importing this module
(and therefore each thin wrapper) never fails when a dependency is missing -- the
caller gets a clean error result instead of an ImportError.
"""
import os
import re
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import outcome
import artifact_store
import consolidation_meta
from compare_core import is_formula_injection   # shared formula-injection guard
from events import ConsolidateResult, Events

# Pull the route token out of "<prefix>_route_<ROUTE>.xlsx".
ROUTE_FROM_NAME = re.compile(r"_route_(\w+)\.xlsx$", re.IGNORECASE)


def _safe_cell(ws, value):
    """A WriteOnlyCell forced to text when `value` would be read by Excel as a
    formula (injection guard); the plain value otherwise (openpyxl handles it).
    The value is preserved exactly — only formula-looking text is neutralized."""
    if is_formula_injection(value):
        c = WriteOnlyCell(ws, value=value)
        c.data_type = "s"
        return c
    return value


def _extract_route(path):
    m = ROUTE_FROM_NAME.search(path.name)
    return m.group(1).upper() if m else path.stem


def _read_header(path, sheet_name):
    """Return row 1 of `sheet_name`, or None if the file isn't shaped like this
    report's export (sheet missing or empty)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return list(row) if row else None
    finally:
        wb.close()


def _stream_data_rows(path, sheet_name):
    """Yield each data row (row 2 onwards) of `sheet_name` as a tuple."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield row
    finally:
        wb.close()


def consolidate_xlsx(*, input_dir, out_path, sheet_name, report_name, title,
                     events=None, confirm_overwrite=None, existed_at_confirm=None,
                     header_override=None, header_comment=None,
                     decorate_workbook=None, commit_guard=None,
                     input_files=None):
    """Combine every per-route XLSX in `input_dir` (reading worksheet
    `sheet_name`) into one workbook at `out_path`, prepending a "Route" column so
    rows from different routes stay distinguishable.

    input_dir / out_path are pathlib.Path. report_name is the friendly name used
    in user-facing messages (UI-neutral). title is the banner shown in the log.

    Optional (used by the Highway Log consolidator to ship CORRECTED column
    labels even though the vendor Excel header is wrong; the data is relabeled by
    POSITION so the rows are untouched):
      * header_override   — list of column labels written instead of the
        file-locked header (must be the same length, else ignored with a note).
      * header_comment     — callable(label) -> openpyxl Comment | None, attached
        to each written header cell (a hover tooltip).
      * decorate_workbook  — callable(wb) run after the rows, before save (e.g.
        to append a Legend sheet).

    ``input_files`` is an optional exact manifest for an attempt-scoped producer.
    When supplied, only those direct children of ``input_dir`` are read; the
    directory is not globbed. Ordinary per-route consolidators omit it and keep
    their existing ``*.xlsx`` discovery behavior.
    """
    events = events or Events()
    input_dir = Path(input_dir)
    out_path = Path(out_path)
    if not _DEPS_OK:
        return ConsolidateResult(
            status="error",
            message="Required components are missing (openpyxl).",
        )
    confirm = confirm_overwrite or (lambda _p: True)

    if not input_dir.exists():
        return ConsolidateResult(
            status="error",
            message=(f"The {report_name} output folder doesn't exist yet:\n{input_dir}\n\n"
                     f"Export the {report_name} report first, then consolidate."),
        )

    # Skip Excel owner-lock stubs (~$foo.xlsx appears the moment a per-route
    # export is open in Excel): they are not workbooks, and counting one as an
    # unreadable input falsely demoted the whole consolidation to PARTIAL.
    if input_files is None:
        files = sorted(p for p in input_dir.glob("*.xlsx")
                       if not p.name.startswith("~$"))
    else:
        root = Path(os.path.abspath(input_dir))
        files = []
        seen = set()
        for member in input_files:
            candidate = Path(os.path.abspath(Path(member)))
            if (candidate.parent != root or candidate.suffix.lower() != ".xlsx"
                    or candidate.name.startswith("~$") or candidate in seen):
                return ConsolidateResult(
                    status="error",
                    message=(f"The {report_name} conversion member manifest was "
                             "invalid; nothing was combined."),
                )
            seen.add(candidate)
            files.append(candidate)
        files.sort(key=lambda p: p.name)
    if not files:
        return ConsolidateResult(
            status="error",
            message=(f"No {report_name} files were found in:\n{input_dir}\n\n"
                     f"Export the {report_name} report first, then consolidate."),
        )

    # Confirm overwrite before reading any inputs. A caller that ALREADY prompted
    # (a per-route converter that confirmed before its long parse) passes the
    # existence it saw at THAT prompt via `existed_at_confirm`; we then skip the
    # initial prompt and only run the pre-replace re-check below (atomic_save_if)
    # with the real confirm — so a late appearance is still caught at the final
    # replace WITHOUT a double prompt for the already-confirmed pre-existing case.
    if existed_at_confirm is None:
        existed_at_confirm = out_path.exists()
        if existed_at_confirm and not confirm(out_path):
            return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"{title} - {len(files)} file(s)")
    events.on_log("=" * 60)
    events.on_log("")

    # Lock in the first readable file's header as the canonical layout. Files that
    # disagree are skipped so misaligned columns can't silently corrupt the
    # combined workbook.
    canonical_header = None
    canonical_source = None
    for p in files:
        try:
            h = _read_header(p, sheet_name)
        except Exception as e:
            events.on_log(f"  {p.name}: header read FAILED ({type(e).__name__}); skipping")
            continue
        if h is None:
            events.on_log(f"  {p.name}: sheet '{sheet_name}' missing; skipping")
            continue
        canonical_header = h
        canonical_source = p.name
        break

    if canonical_header is None:
        return ConsolidateResult(
            status="error",
            message=f"No readable {report_name} files were found in:\n{input_dir}.")

    if not consolidation_meta.guard_allows(commit_guard, out_path):
        return ConsolidateResult(
            status="error",
            message="The destination changed while consolidating; nothing was published.")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not consolidation_meta.guard_allows(commit_guard, out_path):
        return ConsolidateResult(
            status="error",
            message="The destination changed while consolidating; nothing was published.")

    # Style objects are built here (not at module scope) so importing this module
    # never touches openpyxl -- the GUI can import it even if a dep is missing and
    # get a clean error result instead of an ImportError.
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # write_only mode streams rows to disk so the consolidated file can be
    # hundreds of thousands of rows without exhausting memory.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "B2"

    ws.column_dimensions["A"].width = 9
    for i, _ in enumerate(canonical_header, start=2):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Relabel the output header by POSITION when an override is supplied (the
    # data rows are untouched — only the header text changes). Length must match
    # so a positional relabel can't silently shift columns.
    out_header = canonical_header
    if header_override is not None:
        if len(header_override) == len(canonical_header):
            out_header = list(header_override)
        else:
            events.on_log(f"  note: header override has {len(header_override)} "
                          f"columns but the files have {len(canonical_header)}; "
                          "keeping the files' own header.")
    header_values = ["Route"] + [v if v is not None else "" for v in out_header]
    header_cells = []
    for v in header_values:
        cell = WriteOnlyCell(ws, value=str(v))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        if header_comment is not None:
            cm = header_comment(str(v))
            if cm is not None:
                cell.comment = cm
        header_cells.append(cell)
    ws.append(header_cells)

    used_files = 0
    appended_rows = 0
    skipped = []
    failed = []

    for i, p in enumerate(files, 1):
        if events.is_cancelled():
            ws.close()  # finalize the write_only lazy writer (closes its temp file)
            wb.close()  # then release the archive
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i:>3}/{len(files)}] {p.name}"
        try:
            h = _read_header(p, sheet_name)
        except Exception as e:
            events.on_log(f"{prefix} FAILED reading header ({type(e).__name__}): {e}")
            failed.append(p.name)
            continue
        if h is None:
            events.on_log(f"{prefix} skipped: sheet '{sheet_name}' missing")
            skipped.append(p.name)
            continue
        if h != canonical_header:
            events.on_log(f"{prefix} skipped: header differs from {canonical_source}")
            skipped.append(p.name)
            continue

        route = _extract_route(p)
        try:
            count = 0
            for row in _stream_data_rows(p, sheet_name):
                ws.append([route] + [_safe_cell(ws, v) for v in row])
                count += 1
            events.on_log(f"{prefix} +{count} rows  (route {route})")
            appended_rows += count
            used_files += 1
        except Exception as e:
            events.on_log(f"{prefix} FAILED reading rows ({type(e).__name__}): {e}")
            failed.append(p.name)

    # Nothing combined (every file failed/skipped past the header probe) → do
    # NOT save: a header-only workbook claiming "ok" would silently overwrite a
    # good prior consolidation. Report it as an error and leave the file alone.
    if used_files == 0:
        ws.close()
        wb.close()
        return ConsolidateResult(
            status="error",
            message=(f"None of the {len(files)} {report_name} file(s) could be "
                     f"read ({len(skipped)} skipped, {len(failed)} failed) — "
                     f"nothing was combined and the existing file (if any) was "
                     f"left unchanged.\nSee the log for the per-file reasons."))

    # Optional final touch (e.g. append a Legend sheet) before saving.
    if decorate_workbook is not None:
        decorate_workbook(wb)

    events.on_log("")
    events.on_log("Writing consolidated workbook...")
    try:
        # F9: write to a temp sibling + os.replace, so an interrupted/failed write
        # never truncates a prior good consolidated workbook (the destination open in
        # Excel still surfaces as PermissionError, handled below). P12 TOCTOU: the
        # replace is GATED on a re-check (atomic_save_if) — the workbook is already
        # serialized to the temp, so if the destination APPEARED while we combined
        # inputs we ask before overwriting it, without abandoning a half-streamed wb.
        committed = artifact_store.atomic_save_if(
            wb, out_path,
            lambda: (consolidation_meta.guard_allows(commit_guard, out_path)
                     and artifact_store.confirm_late_overwrite(
                         out_path, existed_at_confirm, confirm)))
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not save {out_path.name}.\n\n"
                     "The file is probably open in Excel. Close it and try again.\n"
                     "(Your exported files were not changed.)"),
        )
    if not committed:
        return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    # Skipped/failed inputs mean the combined workbook is INCOMPLETE — lead with a
    # loud warning so a partial result is never mistaken for a full one. P1 makes
    # this a PRODUCER-OWNED outcome: completion=partial with the structured counts
    # (below), so consumers (the matrix, auto-consolidate, the completion card) can
    # flag it WITHOUT parsing this warning text. status stays "ok" so the file that
    # WAS produced is still offered.
    incomplete = bool(skipped or failed)
    summary_lines = []
    if incomplete:
        left_out = len(skipped) + len(failed)
        summary_lines.append(
            f"⚠ INCOMPLETE — {left_out} file(s) were left OUT "
            f"({len(skipped)} skipped, {len(failed)} failed); the combined "
            f"workbook does NOT contain their routes. See the per-file reasons "
            f"above and re-export/repair them before relying on it.")
    summary_lines += [
        f"Files combined: {used_files}",
        f"Rows added:     {appended_rows}",
        f"Files skipped:  {len(skipped)} {skipped if skipped else ''}",
        f"Files failed:   {len(failed)} {failed if failed else ''}",
        f"Output file:    {out_path}",
    ]
    return ConsolidateResult(status="ok", output_path=str(out_path),
                             summary_lines=summary_lines,
                             completion=outcome.PARTIAL if incomplete else outcome.COMPLETE,
                             skipped_inputs=len(skipped), failed_inputs=len(failed))
