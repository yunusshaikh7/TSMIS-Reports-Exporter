"""Shared familiar-layout renderer for the AGGREGATE summary comparisons.

The two SUMMARY reports — Ramp Summary and Intersection Summary — compare a
single statewide category-count table per side (TSMIS vs TSN). compare_core
renders the generic Comparison sheet keyed on the category; this module adds the
FAMILIAR sheet the user knows from the source document: the same sections, in the
same order, with the same labels, now carrying both sides' counts and the
difference. It is plugged in through CompareSchema.extra_sheet_writer (opt-in, so
every non-summary comparison is byte-identical).

One renderer, two reports: each report supplies a SummarySpec (its ordered
sections + the category code/label/slug for each). The Ramp Summary spec lives
here (RAMP_SUMMARY_SPEC); Intersection Summary adds its own. The category SLUGS
match the per-route consolidator's column slugs, so the comparator can map a
consolidated TSMIS workbook's columns to the same categories the TSN PDF parses.

Streaming-safe: only create_sheet + append are used (the comparison workbook is
written in openpyxl write_only mode), mirroring highway_log_columns.write_legend_sheet.
"""
from dataclasses import dataclass, field

try:
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPX = True
except ImportError:
    _OPX = False

# Section band / header colors (echo the consolidator's Combined-sheet palette).
_TITLE_FILL = "1F3864"
_SECTION_FILL = "0070C0"
_DIFF_FONT = "C00000"           # non-zero Δ
_TAB_COLOR = "00B0F0"


@dataclass(frozen=True)
class Cat:
    """One comparison category: `slug` maps to the consolidator column, `label`
    is the short familiar-block text, `key` is the unique compare key (the value
    shown in the generic Comparison sheet's key column). `code` is the within-block
    code token (letter / number / '+') used to map a parsed row to this category
    when the source is block-walked (Intersection Summary); '' for slug-mapped
    reports (Ramp Summary)."""
    slug: str
    label: str
    key: str
    code: str = ""
    # Which system classifies this category: "both" (compared), "tsmis" (TSMIS-only
    # -> lands in 'Only in TSMIS'), or "tsn" (TSN-only -> 'Only in TSN'). Used for
    # the diverged Intersection Summary CONTROL/INTERSECTION-TYPE codes (user chose
    # one-sided, no crosswalk). Default "both" -> every other report unchanged.
    sides: str = "both"


@dataclass(frozen=True)
class Section:
    name: str
    cats: tuple
    # Alternate header spellings that switch to this block when block-walked.
    # PARSING-only: slugs/keys/labels stay derived from `name`, so accepting a
    # renamed source header never changes any workbook text. (The July-2026 site
    # update renamed MASTARM -> MASTERARM; TSN prints keep the old spelling.)
    aliases: tuple = ()


@dataclass(frozen=True)
class SummarySpec:
    """A summary report's familiar layout: ordered sections, a grand-total
    category, and TSMIS-only footnote categories (shown but not section members)."""
    report: str                       # "Ramp Summary"
    sheet_name: str                   # the familiar sheet's tab name
    title: str
    sections: tuple
    total: Cat = None                 # the grand-total row (e.g. Total Number of Ramps)
    footnotes: tuple = field(default_factory=tuple)   # TSMIS-only extras (e.g. no-linework)
    notes: tuple = field(default_factory=tuple)       # extra note lines on the familiar
    #                                                   sheet (e.g. normalizations applied)

    def categories(self):
        """Every COMPARED category in display order (sections then total) as
        (key, slug). Footnotes are NOT compared — they live only on the familiar
        sheet — so they are excluded here."""
        out = []
        for sec in self.sections:
            out += [(c.key, c.slug) for c in sec.cats]
        if self.total is not None:
            out.append((self.total.key, self.total.slug))
        return out

    def categories_for(self, side):
        """The categories EMITTED on `side` ('tsmis' | 'tsn'): shared categories
        plus that side's own one-sided ones, then the grand total. A category the
        other system doesn't classify is omitted here for `side`, so it lands in
        the comparison's 'Only in …' tab (the user's one-sided choice). For an
        all-'both' spec (Ramp Summary) this equals categories()."""
        out = [(c.key, c.slug) for sec in self.sections for c in sec.cats
               if c.sides in ("both", side)]
        if self.total is not None:
            out.append((self.total.key, self.total.slug))
        return out

    def slug_for_key(self):
        return {k: s for k, s in self.categories()}


# =============================================================================
# Ramp Summary canonical spec (slugs == consolidate_ramp_summary column slugs)
# =============================================================================

def _c(slug, label, key, sides="both"):
    return Cat(slug=slug, label=label, key=key, sides=sides)


RAMP_SUMMARY_SPEC = SummarySpec(
    report="Ramp Summary",
    sheet_name="Summary by Category",
    title="Ramp Summary — TSMIS vs TSN by category",
    sections=(
        Section("Highway Groups", (
            _c("hwy_right",         "R - Right",         "Highway Group: R - Right"),
            _c("hwy_divided",       "D - Divided",       "Highway Group: D - Divided"),
            _c("hwy_undivided",     "U - Undivided",     "Highway Group: U - Undivided"),
            _c("hwy_unconstructed", "X - Unconstructed", "Highway Group: X - Unconstructed"),
            _c("hwy_left",          "L - Left",          "Highway Group: L - Left"),
            _c("hwy_others",        "Others",            "Highway Group: Others"),
        )),
        Section("On/Off Indicator", (
            _c("onoff_on",    "ON - On",    "On/Off: ON - On"),
            _c("onoff_off",   "OFF - Off",  "On/Off: OFF - Off"),
            _c("onoff_other", "OTH - Other","On/Off: OTH - Other"),
        )),
        Section("Population Groups", (
            _c("pop_rural_inside",  "R-RURAL -I INSIDE CITY",  "Population: R-RURAL -I INSIDE CITY"),
            _c("pop_rural_outside", "R-RURAL -O OUTSIDE CITY", "Population: R-RURAL -O OUTSIDE CITY"),
            _c("pop_urban_inside",  "U-URBAN -I INSIDE CITY",  "Population: U-URBAN -I INSIDE CITY"),
            _c("pop_urban_outside", "U-URBAN -O OUTSIDE CITY", "Population: U-URBAN -O OUTSIDE CITY"),
            _c("pop_invalid",       "-INVALID DATA",           "Population: -INVALID DATA"),
        )),
        Section("Ramp Types", (
            _c("ramp_A_frontage",    "A - Frontage Road",          "Ramp Type: A - Frontage Road"),
            _c("ramp_B_collector",   "B - Collector Road",         "Ramp Type: B - Collector Road"),
            _c("ramp_C_connector_L", "C - Connector (Left)",       "Ramp Type: C - Direct or Semi-direct Connector (Left)"),
            _c("ramp_D_diamond",     "D - Diamond Type Ramp",      "Ramp Type: D - Diamond Type Ramp"),
            _c("ramp_E_slip",        "E - Slip Ramp",              "Ramp Type: E - Slip Ramp"),
            _c("ramp_F_connector_R", "F - Connector (Right)",      "Ramp Type: F - Direct or Semi-direct Connector (Right)"),
            _c("ramp_G_loop_left",   "G - Loop (w/Left turn)",     "Ramp Type: G - Loop (w/Left turn)"),
            _c("ramp_H_buttonhook",  "H - Buttonhook Ramp",        "Ramp Type: H - Buttonhook Ramp"),
            _c("ramp_J_scissors",    "J - Scissors",               "Ramp Type: J - Scissors"),
            _c("ramp_K_split",       "K - Split Ramp",             "Ramp Type: K - Split Ramp"),
            _c("ramp_L_loop_noleft", "L - Loop without Left Turn", "Ramp Type: L - Loop without Left Turn"),
            _c("ramp_M_two_way",     "M - Two way Ramp Segment",   "Ramp Type: M - Two way Ramp Segment"),
            _c("ramp_P_dummy_paired","P - Dummy Paired",           "Ramp Type: P - Dummy Paired", sides="tsn"),
            _c("ramp_R_rest_area",   "R - Rest Area, Vista Pt",    "Ramp Type: R - Rest Area, Vista Point, Truck Scale"),
            _c("ramp_V_dummy_volume","V - Dummy, Volume only",     "Ramp Type: V - Dummy, Volume only", sides="tsn"),
            _c("ramp_Z_other",       "Z - Other",                  "Ramp Type: Z - Other"),
        )),
    ),
    total=_c("total_ramps", "Total Number of Ramps", "Total Number of Ramps"),
    footnotes=(
        _c("ramp_points_no_linework", "Ramp Points w/out linework",
           "Ramp Points w/out linework"),
    ),
    notes=(
        "Ramp Types P (Dummy Paired) and V (Dummy, Volume only) are TSN bookkeeping "
        "classes the TSMIS summary doesn't tabulate — they stay one-sided by design. "
        "'Ramp Points w/out linework' is the reverse: a TSMIS-only footnote count "
        "with no TSN category, shown below the table and never compared.",
    ),
)


# =============================================================================
# Intersection Summary canonical spec (UNION of the TSN + TSMIS taxonomies)
# =============================================================================
# 11 category blocks; the comparison keys on (block, code-letter) because TSMIS
# reworded many labels ("STOP SIGN"->"STOP SIGNS", "FOUR-WAY"->"4-WAY"). CONTROL
# TYPES: the TSN signal sub-types J–P fold into the shared "Signalized" (S) category
# (see _CONTROL_SIGNAL_FOLD; matches the Detail crosswalk). The codes the TSN summary
# genuinely doesn't tabulate (CONTROL R/O/Q; INTERSECTION TYPE R/C/P; left-chan Y)
# stay one-sided; the "+ no data" buckets the TSN PDF reports as 0 are compared.
# See docs/tsn-parsers.md.
import re as _re


def _bslug(name):
    return _re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def _codeslug(code):
    return "plus" if code == "+" else code.lower().replace("-", "_")


# CONTROL TYPES — signal crosswalk (matches compare_intersection_detail_tsn and the
# TSNR/MIRE reference): TSN records signalized under the legacy sub-types J–P; TSMIS
# stores one code S. counts_from_rows folds J–P -> S within the CONTROL TYPES block
# (for BOTH sides, since they share the mapper), so the summary compares one shared
# "Signalized" category instead of splitting it one-sided. (User, 2026-06-24.)
_IS_CONTROL_TYPES = "CONTROL TYPES"
_CONTROL_SIGNAL_FOLD = frozenset("JKLMNP")                   # -> "S" within CONTROL TYPES

# Codes only ONE system's SUMMARY tabulates (confirmed on the 6.19 raw). With the
# signal fold above there are no TSN-only control codes left. The remaining TSMIS-only
# codes are genuinely absent from the TSN statewide summary PDF (e.g. it has no
# roundabout row, no "Y - channelization not specified" left-chan row). Everything
# else — including the "+ no data" buckets the TSN PDF reports as 0 — is shared and
# compared.
_IS_TSN_ONLY = set()
_IS_TSMIS_ONLY = {
    ("CONTROL TYPES", "R"), ("CONTROL TYPES", "O"), ("CONTROL TYPES", "Q"),
    ("INTERSECTION TYPE", "R"), ("INTERSECTION TYPE", "C"),
    ("INTERSECTION TYPE", "P"), ("INTERSECTION TYPE", "+"),   # new intersection types
    ("MAINLINE LEFT CHANNELIZATION", "Y"),                    # TSN summary has no "Y" row
}


def _icat(block, code, label):
    """One Intersection-Summary category in `block` with within-block `code`."""
    if code.isdigit():
        disp, key = f"{code} lane(s)", f"{block}: {code} lanes"
    else:
        disp, key = f"{code} - {label}", f"{block}: {code} - {label}"
    sides = ("tsn" if (block, code) in _IS_TSN_ONLY
             else "tsmis" if (block, code) in _IS_TSMIS_ONLY else "both")
    return Cat(slug=f"is_{_bslug(block)}_{_codeslug(code)}", label=disp, key=key,
               code=code, sides=sides)


def _isec(block, *cat_specs, aliases=()):
    return Section(block, tuple(_icat(block, c, l) for c, l in cat_specs),
                   aliases=aliases)


_IS_RURAL_URBAN = "RURAL/URBAN/SUBURBAN"

INTERSECTION_SUMMARY_SPEC = SummarySpec(
    report="Intersection Summary",
    sheet_name="Summary by Category",
    title="Intersection Summary — TSMIS vs TSN by category",
    sections=(
        _isec("HIGHWAY GROUP",
              ("R", "RIGHT IND ALIGN"), ("L", "LEFT IND ALIGN"),
              ("X", "UNCONSTRUCTED"), ("U", "UNDIVIDED"), ("D", "DIVIDED")),
        # Rural/Urban: the two '-O OUTSIDE CITY' rows are disambiguated by their
        # R-RURAL / U-URBAN parent (codes R-O / U-O); handled in counts_from_rows.
        Section(_IS_RURAL_URBAN, (
            _icat(_IS_RURAL_URBAN, "R", "RURAL -I INSIDE CITY"),
            _icat(_IS_RURAL_URBAN, "R-O", "RURAL -O OUTSIDE CITY"),
            _icat(_IS_RURAL_URBAN, "U", "URBAN -I INSIDE CITY"),
            _icat(_IS_RURAL_URBAN, "U-O", "URBAN -O OUTSIDE CITY"),
            _icat(_IS_RURAL_URBAN, "+", "INVALID DATA"),
        )),
        _isec("INTERSECTION TYPE",
              ("F", "FOUR-LEGGED"), ("M", "MULTI-LEGGED"), ("S", "OFFSET"),
              ("T", "TEE"), ("Y", "WYE"), ("R", "ROUNDABOUT"),
              ("C", "OTHER CIRCULAR INTERSECTION"), ("P", "MIDBLOCK PED CROSSING (AT GRADE)"),
              ("Z", "OTHER"), ("+", "NO DATA GIVEN")),
        _isec("LIGHTING TYPE",
              ("N", "NO LIGHTING"), ("Y", "LIGHTING"), ("+", "NO DATA GIVEN")),
        # Signal sub-types J–P fold into "S - SIGNALIZED" (see _CONTROL_SIGNAL_FOLD),
        # so they are NOT separate rows here — the single Signalized category carries
        # TSMIS S vs TSN (J–P summed).
        _isec("CONTROL TYPES",
              ("A", "NO CONTROL"), ("B", "STOP SIGNS ON CROSS ST ONLY"),
              ("C", "STOP SIGNS ON MAINLINE ONLY"), ("D", "FOUR-WAY STOP SIGNS"),
              ("E", "4-WAY FLASHER (RED/CROSS ST)"), ("F", "4-WAY FLASHER (RED/MAINLINE)"),
              ("G", "4-WAY FLASHER (RED ON ALL)"), ("H", "YIELD SIGNS (CROSS ST ONLY)"),
              ("I", "YIELD SIGNS (MAIN LINE ONLY)"),
              ("R", "YIELD ALL WAYS (ROUNDABOUT)"), ("S", "SIGNALIZED (incl. TSN J-P)"),
              ("O", "PEDESTRIAN HYBRID BEACON"), ("Q", "FLASH BEACON"),
              ("Z", "OTHER"), ("+", "NO DATA GIVEN")),
        _isec("MAINLINE NUM OF LANES",
              ("1", ""), ("2", ""), ("3", ""), ("4", ""), ("5", ""),
              ("6", ""), ("7", ""), ("8", ""), ("+", "NO DATA GIVEN")),
        # The July-2026 site update renamed this block's header MASTARM ->
        # MASTERARM in the per-route export; the TSN statewide print keeps
        # MASTARM. Alias accepts both; all output text stays "MASTARM".
        _isec("MAINLINE MASTARM",
              ("Y", "YES"), ("N", "NO"), ("+", "NO DATA GIVEN"),
              aliases=("MAINLINE MASTERARM",)),
        _isec("MAINLINE LEFT CHANNELIZATION",
              ("C", "CURBED MEDIAN LEFT TURN CHAN"), ("N", "NO LEFT TURN CHANNELIZATION"),
              ("P", "PAINTED LEFT TURN CHAN"), ("R", "RAISED BARS LEFT TURN CHAN"),
              ("Y", "CHANNELIZATION NOT SPECIFIED"), ("+", "NO DATA GIVEN")),
        _isec("MAINLINE RIGHT CHANNELIZATION",
              ("Y", "FREE RIGHT TURNS"), ("N", "NO FREE RIGHT TURNS"), ("+", "NO DATA GIVEN")),
        _isec("MAINLINE TRAFFIC FLOW",
              ("N", "2 WAY - NO LEFT TURNS"), ("P", "2 WAY WITH LEFT TURN"),
              ("R", "2 WAY - LEFT TURN RESTRICT"), ("W", "ONE WAY TRAFFIC"),
              ("Z", "OTHERS"), ("+", "NO DATA GIVEN")),
    ),
    total=Cat("total_intersections", "Total Intersections", "Total Intersections"),
    notes=(
        "Control Type — the TSN signal sub-types J–P (pretimed / semi- / full-actuated) "
        "are folded into the single 'S - SIGNALIZED' category (matching the Detail "
        "comparison and the TSNR/MIRE reference), so the signalized count compares "
        "directly: TSMIS S vs TSN (J–P summed). Roundabout (R) stays one-sided — the TSN "
        "statewide summary has no roundabout row.",
    ),
)


# =============================================================================
# Strict count parsing + the censused partition contract (CMP-AUD-020/021)
# =============================================================================

def parse_count(value, *, source, category):
    """The ONE strict count parser for every aggregate-summary read path
    (CMP-AUD-021). A category count must be a non-negative whole number: a
    plain int, an integral float (how openpyxl surfaces some numerics), or an
    integer string (commas allowed). Booleans, fractional numbers, and any
    other text raise ValueError naming the source and category — malformed
    data must never silently coerce into a different count."""
    if isinstance(value, bool):
        raise ValueError(f"{source}: the {category!r} count is a boolean "
                         f"({value!r}), not a whole number")
    if isinstance(value, int):
        n = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"{source}: the {category!r} count {value!r} is "
                             "fractional — counts must be whole numbers")
        n = int(value)
    elif isinstance(value, str):
        s = value.strip().replace(",", "")
        if not s:
            raise ValueError(f"{source}: the {category!r} count cell is empty text")
        try:
            n = int(s, 10)
        except ValueError:
            raise ValueError(f"{source}: the {category!r} count {value!r} is "
                             "not a whole number") from None
    else:
        raise ValueError(f"{source}: the {category!r} count has unsupported "
                         f"type {type(value).__name__}")
    if n < 0:
        raise ValueError(f"{source}: the {category!r} count {n} is negative")
    return n


@dataclass(frozen=True)
class SectionRule:
    """The censused partition contract for ONE section on ONE side
    (CMP-AUD-020). `mode`:

      * "exact"   — the side's categories in this section must sum exactly to
                    the grand total;
      * "bounded" — the section may sum BELOW the total only; the censused
                    `reason` names the source classes the side genuinely does
                    not tabulate, and the residual is EXPOSED as a note —
                    never fabricated into any category.

    `extra_slugs` are non-category slugs ADDED to the section sum (the Ramp
    Summary no-linework footnote participates in two TSMIS partitions)."""
    section: str
    mode: str
    reason: str = ""
    extra_slugs: tuple = ()


def reconcile_counts(spec, counts, side, rules, *, source, side_label):
    """Independently validate one side's {slug: count} table against `spec`
    and that side's censused SectionRule contract (CMP-AUD-020). Verifies:

      * every consulted count is a strict non-negative int (defense in depth
        over the loaders' parse_count);
      * the grand total and every side-applicable category are PRESENT — a
        missing source fact is a hard stop, never a fabricated zero;
      * every section partitions the grand total per its rule.

    Returns the note lines exposing bounded residuals (empty when every block
    is exact); raises ValueError (naming `source`) on any violation."""
    rules_by_name = {}
    for rule in rules:
        if rule.section in rules_by_name:
            raise ValueError(f"{source}: duplicate SectionRule for {rule.section!r}")
        rules_by_name[rule.section] = rule
    section_names = {s.name for s in spec.sections}
    missing_rules = [s.name for s in spec.sections if s.name not in rules_by_name]
    unknown_rules = [n for n in rules_by_name if n not in section_names]
    if missing_rules or unknown_rules:
        raise ValueError(f"{source}: the SectionRule contract does not match "
                         f"the spec (missing {missing_rules}, unknown "
                         f"{unknown_rules})")

    def count_of(slug, label):
        v = counts.get(slug)
        if v is None:
            return None
        if isinstance(v, bool) or not isinstance(v, int) or v < 0:
            raise ValueError(f"{source}: the {label!r} count {v!r} is not a "
                             "non-negative whole number")
        return v

    if spec.total is None:
        raise ValueError(f"{source}: the report spec has no grand-total category")
    total = count_of(spec.total.slug, spec.total.key)
    if total is None:
        raise ValueError(f"{source}: the grand total ({spec.total.key}) was "
                         "not found — cannot validate the category table")

    missing = [c.key for sec in spec.sections for c in sec.cats
               if c.sides in ("both", side) and count_of(c.slug, c.key) is None]
    if missing:
        raise ValueError(
            f"{source}: {len(missing)} expected categor"
            f"{'y is' if len(missing) == 1 else 'ies are'} missing: "
            + ", ".join(missing[:6]) + ("…" if len(missing) > 6 else ""))

    notes = []
    for sec in spec.sections:
        rule = rules_by_name[sec.name]
        ssum = sum(count_of(c.slug, c.key) for c in sec.cats
                   if c.sides in ("both", side))
        for slug in rule.extra_slugs:
            extra = count_of(slug, slug)
            if extra is None:
                raise ValueError(f"{source}: the '{sec.name}' partition needs "
                                 f"{slug!r}, which is missing")
            ssum += extra
        if rule.mode == "exact":
            if ssum != total:
                raise ValueError(
                    f"{source}: the '{sec.name}' block sums {ssum:,} of "
                    f"{total:,} — the source layout may have changed (renamed "
                    "header, new code, or dropped rows); refusing to compare "
                    "a table that does not reconcile")
        elif rule.mode == "bounded":
            if ssum > total:
                raise ValueError(
                    f"{source}: the '{sec.name}' block sums {ssum:,}, MORE "
                    f"than the total {total:,} — the table does not reconcile")
            if ssum < total:
                notes.append(f"{side_label} '{sec.name}': {ssum:,} of {total:,} "
                             f"tabulated ({total - ssum:,} not — {rule.reason}).")
        else:
            raise ValueError(f"{source}: unknown SectionRule mode {rule.mode!r} "
                             f"for {sec.name!r}")
    return notes


# =============================================================================
# Spec-driven block-walk: a (count, text) row sequence -> {slug: count}
# Shared by the TSN PDF parser and the TSMIS per-route consolidator (both feed a
# (count_or_None, code-text) stream; block headers switch the active block).
# =============================================================================
_HEADER_NOISE = _re.compile(r"[<>]|-{2,}")


def _norm_header(text):
    return _re.sub(r"\s+", " ", _HEADER_NOISE.sub(" ", str(text or ""))).strip().upper()


def _plain_code(text):
    """The within-block code token from a row's code text: a leading 'X-' letter,
    a leading number (lanes), or '+'. None for a non-data row."""
    t = str(text or "").strip()
    if not t:
        return None
    if t[0] == "+":
        return "+"
    m = _re.match(r"^([A-Za-z])\s*-", t)
    if m:
        return m.group(1).upper()
    m = _re.match(r"^(\d+)\b", t)
    if m:
        return m.group(1)
    return None


def counts_from_rows(spec, rows, *, source="source rows"):
    """Map a (count_or_None, text) row stream to {slug: count} using `spec`'s
    block structure. A row whose text matches a block header switches the active
    block; data rows (numeric count) map by within-block code. The Rural/Urban
    block's two '-O OUTSIDE CITY' rows are bound to their R-RURAL / U-URBAN
    parent; the parent updates from the LABEL even on a count-less row, and a
    count-carrying '-O' with NO parent in the block is an error (CMP-AUD-023 —
    it must never silently default to Rural). Counts are strict (parse_count)."""
    headers = {_norm_header(s.name): s for s in spec.sections}
    for s in spec.sections:                      # renamed-header aliases (parse-only)
        for alias in s.aliases:
            headers[_norm_header(alias)] = s
    by_block = {s.name: {c.code: c for c in s.cats} for s in spec.sections}
    out, cur, ru_parent = {}, None, None
    for count, text in rows:
        t = str(text or "").strip()
        h = _norm_header(t)
        if h in headers:
            cur, ru_parent = headers[h], None
            continue
        if cur is None:
            continue
        if cur.name == _IS_RURAL_URBAN:
            up = t.upper()
            # Parent context comes from the LABEL, before any numeric gate: a
            # count-less R-RURAL/U-URBAN row still names the parent its
            # following '-O OUTSIDE CITY' row belongs to (CMP-AUD-023).
            if up.startswith("R-RURAL"):
                ru_parent, code = "R", "R"
            elif up.startswith("U-URBAN"):
                ru_parent, code = "U", "U"
            elif up.startswith("-O"):
                if ru_parent is None and count is not None:
                    raise ValueError(
                        f"{source}: an '-O OUTSIDE CITY' row (count {count!r}) "
                        "has no preceding R-RURAL/U-URBAN parent — the count "
                        "cannot be attributed")
                code = f"{ru_parent}-O" if ru_parent is not None else None
            elif up.startswith("+"):
                code = "+"
            else:
                code = None
        else:
            code = _plain_code(t)
            if cur.name == _IS_CONTROL_TYPES and code in _CONTROL_SIGNAL_FOLD:
                code = "S"                       # fold TSN signal sub-types J–P -> S
        if count is None:
            continue
        cat = by_block[cur.name].get(code) if code is not None else None
        if cat is not None:
            out[cat.slug] = out.get(cat.slug, 0) + parse_count(
                count, source=source, category=f"{cur.name}: {t}")
    return out


# =============================================================================
# Familiar-sheet renderer (write_only-safe; plugged in via extra_sheet_writer)
# =============================================================================

def _as_int(v):
    """A count cell to int, or None when it isn't a plain number."""
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    s = str(v or "").strip().replace(",", "")
    try:
        return int(s)
    except ValueError:
        return None


def make_extra_sheet_writer(spec, footnote_values=None, extra_notes=None):
    """Return an extra_sheet_writer(wb, ctx) that appends `spec`'s familiar
    category-comparison sheet. ctx carries rows_a/rows_b (each [key, count]) and
    the side labels — see compare_core.CompareSchema.extra_sheet_writer.

    `footnote_values` (CMP-AUD-024, opt-in): a {footnote.key: value} mapping supplied
    OUT OF BAND — footnotes are display-only and are deliberately NOT in the compared
    rows, so a footnote can never become a one-sided comparison row or move the verdict.
    The comparator binds a fresh mapping per run and the loader populates it before the
    writer runs. When omitted, footnote display falls back to the compared rows (legacy).

    `extra_notes` (CMP-AUD-020, opt-in): a per-run list of note LINES the loader
    fills (the censused bounded-partition residuals, e.g. the 22 TSMIS ramps in
    TSN-only P/V classes). Display-only exposure on the familiar sheet — never a
    warning (warnings mean unreadable inputs and would mark the run incomplete)."""
    def writer(wb, ctx):
        if not _OPX:
            return None
        return _render(wb, ctx, spec, footnote_values=footnote_values,
                       extra_notes=extra_notes)
    return writer


def _render(wb, ctx, spec, footnote_values=None, extra_notes=None):
    sc = ctx["sc"]
    side_a, side_b = sc.side_a, sc.side_b          # the side LABELS ("TSMIS"/"TSN")
    file_a, file_b = ctx.get("side_a", ""), ctx.get("side_b", "")   # the source filenames
    va = {r[0]: _as_int(r[1]) for r in ctx["rows_a"]}
    vb = {r[0]: _as_int(r[1]) for r in ctx["rows_b"]}

    ws = wb.create_sheet(spec.sheet_name)
    ws.sheet_properties.tabColor = _TAB_COLOR
    write_only = getattr(wb, "write_only", False)

    title_font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", start_color=_TITLE_FILL)
    sec_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    sec_fill = PatternFill("solid", start_color=_SECTION_FILL)
    head_font = Font(name="Arial", size=10, bold=True)
    body = Font(name="Arial", size=10)
    diff_font = Font(name="Arial", size=10, bold=True, color=_DIFF_FONT)
    note_font = Font(name="Arial", size=9, italic=True, color="595959")
    right = Alignment(horizontal="right")
    left = Alignment(horizontal="left")

    def cell(value, font=body, fill=None, align=None):
        if not write_only:
            return value
        c = WriteOnlyCell(ws, value=value)
        c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        return c

    for col, w in (("A", 34), ("B", 13), ("C", 13), ("D", 10)):
        ws.column_dimensions[col].width = w

    ws.append([cell(spec.title, title_font, title_fill)])
    ws.append([cell(f"Counts per category. Δ = {side_b} − {side_a}; a non-zero Δ is "
                    "flagged. Categories one system doesn't classify show 0 on that "
                    "side (e.g. TSN-only ramp types P / V).", note_font)])
    if file_a or file_b:
        ws.append([cell(f"{side_a} = {file_a}    {side_b} = {file_b}", note_font)])
    for n in spec.notes:
        ws.append([cell(n, note_font)])
    for n in (extra_notes or ()):
        ws.append([cell(str(n), note_font)])
    ws.append([])
    ws.append([cell("Category", head_font), cell(side_a, head_font, align=right),
               cell(side_b, head_font, align=right), cell("Δ", head_font, align=right)])

    def value_row(label, key):
        a, b = va.get(key), vb.get(key)
        delta = (b - a) if (a is not None and b is not None) else None
        differ = delta is not None and delta != 0
        f = diff_font if differ else body
        return [cell(label, body, align=left),
                cell(a, f, align=right), cell(b, f, align=right),
                cell(delta if delta is not None else "", f, align=right)]

    for sec in spec.sections:
        ws.append([cell(sec.name, sec_font, sec_fill),
                   cell("", sec_font, sec_fill), cell("", sec_font, sec_fill),
                   cell("", sec_font, sec_fill)])
        for c in sec.cats:
            ws.append(value_row(c.label, c.key))

    if spec.total is not None:
        ws.append([])
        a, b = va.get(spec.total.key), vb.get(spec.total.key)
        delta = (b - a) if (a is not None and b is not None) else None
        f = diff_font if (delta is not None and delta != 0) else head_font
        ws.append([cell(spec.total.label, head_font, align=left),
                   cell(a, f, align=right), cell(b, f, align=right),
                   cell(delta if delta is not None else "", f, align=right)])

    if spec.footnotes:
        ws.append([])
        ws.append([cell(f"Reported by {side_a} only (not a {side_b} category):", note_font)])
        for fnote in spec.footnotes:
            # Footnote values come out of band (CMP-AUD-024) so they never enter the
            # compared universe; fall back to the compared row only for legacy callers.
            a = (footnote_values.get(fnote.key) if footnote_values is not None
                 else va.get(fnote.key))
            ws.append([cell(fnote.label, body, align=left),
                       cell(a, body, align=right), cell("", body), cell("", body)])
    return ws
