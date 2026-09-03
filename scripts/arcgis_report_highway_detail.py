"""Build OUR Highway Detail report from the ArcGIS layer library.

This report has its OWN build (v0.39.2). It shares the span engine with the
Clean Road build but not that build's output, because the two are reproducing
different things: Clean Road reproduces the vendor's 74-column `CA HIGHWAYS`
table, so it segments on TSN's columns and dates its blocks the way TSN's table
was measured. Highway Detail is a REPORT, and gets its own segmentation and its
own measured rules. Two passes:

  1. the shared span engine over `SEGMENT_TAGS` — the 19 layers this report
     actually prints — with `primary_eff` block dates;
  2. the projection of that table onto the report's 34 printed columns.

Building it off the Clean Road output instead cost twice, measured against the
real 2026-08-17 export: ~6,500 boundaries the report never draws (splits on
ADT/terrain/design-speed/toll/forest, undone afterwards by the merge pass and
only where every printed column happened to agree), and a block effective date
computed as the OLDEST of five member layers, which scored 56-60% where the
report's own rule scores 79-80% (DA1).

The projection is still not a column rename: even on this report's own layers a
boundary can exist where no printed value changes (a re-inventoried span, a
segment-order change), so adjacent spans that agree across every printed column
and touch exactly are MERGED, and the merged span's own end minus begin is the
printed Length. Route 001's first record reads `R000.129 / 000.075` — the
export's own value, which the unmerged span (0.041) does not.

`RU Eff` used to be the one column with nothing to print from (DA2): the CA
HIGHWAYS table gives four attribute blocks an effective date but gives population
only a CODE, so the projection emitted it empty and declared it context. The date
was never actually missing — it is `SHS Population.InventoryItemStartDate`, the
same layer row the code comes from — so the Clean Road build now carries it as
the build-only column `THY_POPULATION_EFF_DATE`
(`clean_highway_columns.BUILD_ONLY_COLUMNS`) and this projection prints it like
any other date. Every column the report prints is now sourced and counted; there
are no context columns left here.

Vintage is the other thing to know: the built workbook is a reconstruction
AS OF a chosen date, so a faithful comparison needs that as-of date to match the
export day being compared. The build's own as-of is read back from its marker
sheet and carried onto ours, and the comparison states both.

Console-free; returns a ConsolidateResult. openpyxl loads lazily.
"""
from __future__ import annotations

import logging
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import clean_highway_columns as chc
import consolidate_clean_highway as cch
import consolidation_meta
import highway_detail_columns as hdc
import outcome
import paths
from events import ConsolidateResult, Events

log = logging.getLogger(__name__)

REPORT_NAME = "Highway Detail (ArcGIS)"
SHEET_NAME = "Highway Detail"          # matches the consolidated export's sheet
MARKER_SHEET = "ArcGIS Report Build"
PROVENANCE_SHEET = "Provenance"
BUILD_VERSION = 1
FILENAME = "highway_detail_from_layers.xlsx"
OUT_DIR = paths.OUTPUT_ROOT / "arcgis_reports"
OUT_PATH = OUT_DIR / FILENAME

ROUTE_COL = hdc.ROUTE_COL
HEADER = [ROUTE_COL] + list(hdc.HEADER)
# The "Reports vs layers" library's two hooks: the sidecar payload key the
# outcome record carries the build facts under, and the layers the build needs
# (the span engine refuses without every highway layer, so they all gate it).
SIDECAR_KEY = "arcgis_report_build"
REQUIRED_LAYERS = cch.HIGHWAY_LAYERS

# --------------------------------------------------------------------------- #
# The projection: report column -> (THY column(s), how). "how" is the shared
# normalization name so the Provenance sheet can state it verbatim; `_project`
# below is the single implementation. Post Mile and Length are SPAN-derived
# (they describe the merged stretch, not one THY cell) and Route is the
# consolidated export's leading identity column.
# --------------------------------------------------------------------------- #
_SPAN = "(span)"
_NONE = ""
PROJECTION = {
    "Post Mile": (("THY_PM_PREFIX_CODE", "THY_BEGIN_PM_AMT",
                   "THY_PM_SUFFIX_CODE", "THY_EQUATE_CODE"), "postmile"),
    "Length": ((_SPAN,), "merged end PM - begin PM, 3 decimals"),
    "Date of Rec": (("THY_RECORD_DATE",), "date"),
    "HG": (("THY_HIGHWAY_GROUP_CODE",), "code"),
    "AC": (("THY_HIGHWAY_ACCESS_CODE",), "code"),
    "Acc-Cont Eff": (("THY_ACCESS_EFF_DATE",), "date"),
    "City": (("THY_CITY_CODE",), "code"),
    "RU": (("THY_POPULATION_CODE",), "code"),
    "RU Eff": (("THY_POPULATION_EFF_DATE",), "date"),
    "Description": (("THY_LANDMARK_SHORT_DESC",), "text, upper-cased"),
    "NA": (("THY_NON_ADD_CODE",), "'A' (add mileage) prints blank"),
    "LB Eff": (("THY_LEFT_ROAD_EFF_DATE",), "date"),
    "LB S/T": (("THY_LT_SURF_TYPE_CODE",), "code"),
    "LB #Ln": (("THY_LT_LANES_AMT",), "2-digit"),
    "LB S/F": (("THY_LT_SPEC_FEATURES_CODE",), "code"),
    "LB OT-TO": (("THY_LT_O_SHD_TOT_WIDTH_AMT",), "2-digit"),
    "LB OT-TR": (("THY_LT_O_SHD_TRT_WIDTH_AMT",), "2-digit"),
    "LB Wid": (("THY_LT_TRAV_WAY_WIDTH_AMT",), "2-digit"),
    "LB IN-TO": (("THY_LT_I_SHD_TOT_WIDTH_AMT",), "2-digit"),
    "LB IN-TR": (("THY_LT_I_SHD_TRT_WIDTH_AMT",), "2-digit"),
    "Med Eff": (("THY_MEDIAN_EFF_DATE",), "date"),
    "Med T": (("THY_MEDIAN_TYPE_CODE",), "code"),
    "Med C": (("THY_CURB_LANDSCAPE_CODE",), "code"),
    "Med B": (("THY_MEDIAN_BARRIER_CODE",), "code"),
    "Med V/WDA": (("THY_MEDIAN_WIDTH_AMT", "THY_MEDIAN_WIDTH_VAR_CODE"),
                  "width + variance glued"),
    "RB Eff": (("THY_RIGHT_ROAD_EFF_DATE",), "date"),
    "RB S/T": (("THY_RT_SURF_TYPE_CODE",), "code"),
    "RB #Ln": (("THY_RT_LANES_AMT",), "2-digit"),
    "RB S/F": (("THY_RT_SPEC_FEATURES_CODE",), "code"),
    "RB IN-TO": (("THY_RT_I_SHD_TOT_WIDTH_AMT",), "2-digit"),
    "RB IN-TR": (("THY_RT_I_SHD_TRT_WIDTH_AMT",), "2-digit"),
    "RB Wid": (("THY_RT_TRAV_WAY_WIDTH_AMT",), "2-digit"),
    "RB OT-TO": (("THY_RT_O_SHD_TOT_WIDTH_AMT",), "2-digit"),
    "RB OT-TR": (("THY_RT_O_SHD_TRT_WIDTH_AMT",), "2-digit"),
}
assert set(PROJECTION) == set(hdc.HEADER), "the projection must cover every column"

# Columns the comparison SHOWS but never COUNTS: the report prints them, but the
# THY build has no source to render them from, so a difference would only ever
# measure our own gap. Kept PRESENT with both sides visible (the clean-road
# CONTEXT_COLUMNS convention).
CONTEXT_COLUMNS = tuple(name for name, (src, _how) in PROJECTION.items()
                        if src == (_NONE,))

# The columns a row boundary is allowed to depend on: everything the report
# prints except its own position/extent AND the description.
#
# Description is START-ANCHORED, not a span value: landmarks are POINT features
# in the layer library, so the record beginning at a landmark's postmile prints
# it and the stretch that follows carries none. Treating a blank description as
# a difference split route 001's first record in two (000.041 + 000.034) where
# the export prints one 000.075 record. So a following span whose description is
# BLANK continues the record and keeps the first one's text; a following span
# with ANY non-blank description starts a new record, because a landmark is
# exactly what does start one — including one whose text repeats the record
# before it (two consecutive "BEGIN OF COUNTY" landmarks are two records).
# Measured (DA6): allowing a repeated description to merge cost 85 records the
# export prints.
_MERGE_FIELDS = tuple(f for f in hdc.HEADER
                      if f not in ("Post Mile", "Length", "Description"))

# The span layers THIS REPORT is built from — and therefore the ones its row
# boundaries are cut on. `crl.overlay` splits a segment wherever ANY supplied
# span changes, so the tag set IS the segmentation.
#
# The Clean Road build supplies all 25 because it is reproducing the TSN CA
# HIGHWAYS table, which segments on its own 74 columns. Highway Detail prints 34
# of them, so building it through that segmentation produced ~6,500 splits the
# report does not draw and needed a merge pass to undo them — which only worked
# where every printed column happened to agree, and left the block effective
# dates fragmenting `Length` wherever they flickered (DA1).
#
# Building on the report's OWN layers puts the boundaries where the report puts
# them, so nothing has to be undone. These are the five the report never prints:
#
#   TVS   ADT / profile / change-per-mile
#   TER   terrain
#   DSP   design speed
#   TOLL  toll type
#   FOR   forest highway
#
# `check_arcgis_report` derives the same set from PROJECTION + the clean-road
# PROVENANCE table and fails if this list drifts from what the report prints.
_NOT_PRINTED = ("TVS", "TER", "DSP", "TOLL", "FOR")
SEGMENT_TAGS = tuple(t for t in cch.SPAN_LAYERS if t not in _NOT_PRINTED)

_T = {name: i for i, name in enumerate(chc.ARC_HEADER)}


# --------------------------------------------------------------------------- #
# cell normalizations (one home; the Provenance sheet names them)
# --------------------------------------------------------------------------- #
def _s(v):
    return "" if v is None else str(v).strip()


def _dec(v):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):  # silent-ok: a pure numeric predicate — a non-numeric cell simply has no number, and every caller renders that as the report's blank
        return None


def _pm3(v):
    """A postmile/length to the report's fixed 3-decimal form ('000.075')."""
    d = _dec(v)
    return "" if d is None else f"{d:07.3f}"


def _date(v):
    """A build date to the report's 'YY-MM-DD' print form."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%y-%m-%d")
    return _s(v)


def _n2(v):
    """A count/width to the report's zero-padded 2-digit form ('03', '00')."""
    s = _s(v)
    if not s:
        return ""
    d = _dec(s)
    return f"{int(d):02d}" if d is not None else s


def _wda(width, variance):
    """Median width + variance glued the way the report prints them ('14Z')."""
    d = _dec(width)
    return "" if d is None else f"{int(d):02d}{_s(variance)}"


def _na(v):
    """Non-add: the report leaves ordinary ADD mileage blank where the table
    carries an explicit 'A'."""
    s = _s(v).upper()
    return "" if s == "A" else s


def _desc(v):
    """The landmark description as the report prints it (upper-case)."""
    return _s(v).upper()


def _at(r, name):
    i = _T[name]
    return r[i] if i < len(r) else None


def _project(r):
    """The printed value of every report column that comes from ONE THY row.
    Post Mile and Length are span-derived and filled by the merge."""
    return {
        "Post Mile": "", "Length": "",
        "Date of Rec": _date(_at(r, "THY_RECORD_DATE")),
        "HG": _s(_at(r, "THY_HIGHWAY_GROUP_CODE")),
        "AC": _s(_at(r, "THY_HIGHWAY_ACCESS_CODE")),
        "Acc-Cont Eff": _date(_at(r, "THY_ACCESS_EFF_DATE")),
        "City": _s(_at(r, "THY_CITY_CODE")),
        "RU": _s(_at(r, "THY_POPULATION_CODE")),
        "RU Eff": _date(_at(r, "THY_POPULATION_EFF_DATE")),
        "Description": _desc(_at(r, "THY_LANDMARK_SHORT_DESC")),
        "NA": _na(_at(r, "THY_NON_ADD_CODE")),
        "LB Eff": _date(_at(r, "THY_LEFT_ROAD_EFF_DATE")),
        "LB S/T": _s(_at(r, "THY_LT_SURF_TYPE_CODE")),
        "LB #Ln": _n2(_at(r, "THY_LT_LANES_AMT")),
        "LB S/F": _s(_at(r, "THY_LT_SPEC_FEATURES_CODE")),
        "LB OT-TO": _n2(_at(r, "THY_LT_O_SHD_TOT_WIDTH_AMT")),
        "LB OT-TR": _n2(_at(r, "THY_LT_O_SHD_TRT_WIDTH_AMT")),
        "LB Wid": _n2(_at(r, "THY_LT_TRAV_WAY_WIDTH_AMT")),
        "LB IN-TO": _n2(_at(r, "THY_LT_I_SHD_TOT_WIDTH_AMT")),
        "LB IN-TR": _n2(_at(r, "THY_LT_I_SHD_TRT_WIDTH_AMT")),
        "Med Eff": _date(_at(r, "THY_MEDIAN_EFF_DATE")),
        "Med T": _s(_at(r, "THY_MEDIAN_TYPE_CODE")),
        "Med C": _s(_at(r, "THY_CURB_LANDSCAPE_CODE")),
        "Med B": _s(_at(r, "THY_MEDIAN_BARRIER_CODE")),
        "Med V/WDA": _wda(_at(r, "THY_MEDIAN_WIDTH_AMT"),
                          _at(r, "THY_MEDIAN_WIDTH_VAR_CODE")),
        "RB Eff": _date(_at(r, "THY_RIGHT_ROAD_EFF_DATE")),
        "RB S/T": _s(_at(r, "THY_RT_SURF_TYPE_CODE")),
        "RB #Ln": _n2(_at(r, "THY_RT_LANES_AMT")),
        "RB S/F": _s(_at(r, "THY_RT_SPEC_FEATURES_CODE")),
        "RB IN-TO": _n2(_at(r, "THY_RT_I_SHD_TOT_WIDTH_AMT")),
        "RB IN-TR": _n2(_at(r, "THY_RT_I_SHD_TRT_WIDTH_AMT")),
        "RB Wid": _n2(_at(r, "THY_RT_TRAV_WAY_WIDTH_AMT")),
        "RB OT-TO": _n2(_at(r, "THY_RT_O_SHD_TOT_WIDTH_AMT")),
        "RB OT-TR": _n2(_at(r, "THY_RT_O_SHD_TRT_WIDTH_AMT")),
    }


def _route_token(r):
    """The consolidated export's route identity ('001', '005S')."""
    return _s(_at(r, "THY_ROUTE_NAME")) + _s(_at(r, "THY_ROUTE_SUFFIX_CODE")).upper()


def _pm_suffix(r):
    """The report's trailing postmile marker: the independent-roadbed R/L, plus
    the equation-point E where the build placed one."""
    suffix = _s(_at(r, "THY_PM_SUFFIX_CODE")).upper()
    equate = _s(_at(r, "THY_EQUATE_CODE")).upper()
    return suffix + ("E" if equate == "E" else "")


# --------------------------------------------------------------------------- #
# reading the built workbook
# --------------------------------------------------------------------------- #
def built_facts(path):
    """The build facts the source workbook records about itself (as-of date,
    build version, layer library). Missing/unreadable facts read as unknown —
    never invented."""
    from openpyxl import load_workbook

    facts = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if chc.ARC_MARKER_SHEET not in wb.sheetnames:
            return facts
        for r in wb[chc.ARC_MARKER_SHEET].iter_rows(values_only=True):
            if not r or r[0] is None:
                continue
            key, val = _s(r[0]), (r[1] if len(r) > 1 else None)
            if key == "As-of date":
                facts["asof"] = _s(val)
            elif key == "Build version":
                facts["source_build_version"] = _s(val)
            elif key == "Layer library":
                facts["layer_library"] = _s(val)
            # The drop the source table was built from — carried onto the
            # report so the library can tell a build from the staged drop
            # apart from one built off an older export.
            elif key == "Layer drop exported":
                facts["drop_exported"] = _s(val)
            elif key == "Layer drop fingerprint":
                facts["drop_fingerprint"] = _s(val)
            # HF-01/RB-1: the source build's unassertable spans. A projection
            # inherits them — the marker token travels into whichever report
            # column the span would have painted — so the counts travel with
            # it. An older, skip-free build simply has no such rows.
            elif key == "Skipped source spans":
                facts["skipped_source_spans"] = _s(val)
            elif key == "Marked anchor cells":
                facts["marked_anchor_cells"] = _s(val)
            elif key == "Unavailable marker":
                facts["unavailable_marker"] = _s(val)
            elif key == "Skipped source reason":
                facts["skipped_source_reason"] = _s(val)
    finally:
        wb.close()
    return facts


def is_arcgis_report(path):
    """Is `path` a workbook THIS module built? Identity is the marker sheet it
    writes — the comparison uses it to enforce which side is which, so a
    missing/unreadable workbook answers False rather than raising."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True)
    except Exception as e:    # silent-ok: a role probe — an unopenable file is simply not one of ours, and the loader that follows reports the real problem
        log.info("arcgis report probe: %s unreadable (%s: %s)",
                 path, type(e).__name__, e)
        return False
    try:
        return MARKER_SHEET in wb.sheetnames
    finally:
        wb.close()


def _skipped_spans(facts):
    """How many source spans the CA HIGHWAYS build could not place, as an int.
    Absent / unreadable / not a number reads as 0 — a build that says nothing
    about skips is a build with none (the pre-HF-01 shape), and a bad value is
    never rounded up into a claim."""
    try:
        return max(0, int(str(facts.get("skipped_source_spans", "")).strip()
                          or 0))
    except ValueError:      # silent-ok: a marker sheet that does not state a number states nothing
        return 0


def report_facts(path):
    """The build facts THIS module's marker sheet records (as-of date, source
    workbook, row counts, and any unassertable-span counts inherited from the
    CA HIGHWAYS build). `{}` when absent/unreadable — never invented."""
    from openpyxl import load_workbook

    facts = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if MARKER_SHEET not in wb.sheetnames:
            return facts
        keys = {"As-of date": "asof", "Source workbook": "source",
                "Build version": "build_version",
                "Layer drop exported": "drop_exported",
                "Layer drop fingerprint": "drop_fingerprint",
                "Report records written": "rows",
                "Rows merged away": "merged_away", "Routes": "routes",
                "Skipped source spans": "skipped_source_spans",
                "Marked anchor cells": "marked_anchor_cells",
                "Unavailable marker": "unavailable_marker",
                "Skipped source reason": "skipped_source_reason"}
        for r in wb[MARKER_SHEET].iter_rows(values_only=True):
            if not r or r[0] is None:
                continue
            key = keys.get(_s(r[0]))
            if key:
                facts[key] = _s(r[1] if len(r) > 1 else None)
        return facts
    finally:
        wb.close()


def _require_built(path):
    """The built CA HIGHWAYS workbook, verified to BE one (its own data sheet
    with the exact 74-column THY header) — never a lookalike."""
    from openpyxl import load_workbook

    path = Path(path)
    if not path.is_file():
        raise ValueError(
            "The Clean Road Highway workbook has not been built yet.\n\n"
            "Build it on the Clean Road tab first — this report is projected "
            "from it, not from the layers directly.")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if chc.ARC_SHEET not in wb.sheetnames:
            raise ValueError(
                f"{path.name} has no '{chc.ARC_SHEET}' sheet — it is not the "
                "ArcGIS Clean Road Highway build.")
        ws = wb[chc.ARC_SHEET]
        it = ws.iter_rows(values_only=True)
        header = [_s(c) for c in (next(it, ()) or ())]
        if header[:len(chc.ARC_HEADER)] != list(chc.ARC_HEADER):
            raise ValueError(
                f"{path.name} does not carry the {len(chc.ARC_HEADER)}-column "
                "CA HIGHWAYS header — rebuild the Clean Road Highway workbook.")
        for row in it:
            if row and any(c is not None for c in row):
                yield row
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# the projection + merge
# --------------------------------------------------------------------------- #
def project_rows(built_rows, events=None):
    """(rows, stats) — the report-shaped rows, newest merge applied.

    Grouping is by (route, county, PM prefix, PM suffix): a row boundary can
    only be dissolved between spans that describe the same postmile line, and
    the county+prefix+suffix triple is what makes two postmiles comparable
    (the same number recurs on a different county or a realignment prefix)."""
    events = events or Events()
    groups = defaultdict(list)
    raw = 0
    for r in built_rows:
        begin, end = _dec(_at(r, "THY_BEGIN_PM_AMT")), _dec(_at(r, "THY_END_PM_AMT"))
        if begin is None or end is None:
            continue          # a span with no readable extent has no printed row
        raw += 1
        key = (_route_token(r), _s(_at(r, "THY_COUNTY_CODE")),
               _s(_at(r, "THY_PM_PREFIX_CODE")).upper(), _pm_suffix(r))
        groups[key].append((begin, end, _project(r)))

    rows, merged_away = [], 0
    for (route, _county, prefix, suffix) in sorted(groups):
        spans = sorted(groups[(route, _county, prefix, suffix)],
                       key=lambda t: (t[0], t[1]))
        runs = []
        for begin, end, proj in spans:
            prev = runs[-1] if runs else None
            desc = proj["Description"]
            if (prev is not None and prev[1] == begin
                    and all(prev[2][f] == proj[f] for f in _MERGE_FIELDS)
                    and desc == ""):
                prev[1] = end                       # same printed record, longer
                merged_away += 1
            else:
                runs.append([begin, end, proj])
        for begin, end, proj in runs:
            proj = dict(proj)
            proj["Post Mile"] = f"{prefix}{begin:07.3f}{suffix}"
            proj["Length"] = _pm3(end - begin)
            rows.append([route] + [proj[c] for c in hdc.HEADER])

    rows.sort(key=lambda r: (r[0], r[1]))
    stats = {"source_rows": raw, "rows": len(rows), "merged_away": merged_away,
             "routes": len({r[0] for r in rows})}
    events.on_log(f"  projected {raw:,} CA HIGHWAYS rows onto the report's "
                  f"{len(hdc.HEADER)} columns -> {len(rows):,} records "
                  f"({merged_away:,} merged away)")
    return rows, stats


# --------------------------------------------------------------------------- #
# workbook
# --------------------------------------------------------------------------- #
def _provenance_rows():
    """(report column, THY column(s), how, tier) for the Provenance sheet —
    every printed column indexed back to what produced it."""
    out = []
    for name in hdc.HEADER:
        src, how = PROJECTION[name]
        if src == (_NONE,):
            tier, cols = "no source", ""
        elif src == (_SPAN,):
            tier, cols = "span-derived", ""
        else:
            tier, cols = "projected", ", ".join(src)
        out.append((name, cols, how, tier))
    return out


def _write_workbook(out_path, rows, stats, facts, source_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    head_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", start_color="1F3864")
    ws.append(HEADER)
    for c in next(ws.iter_rows(min_row=1, max_row=1)):
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    widths = {"Route": 7, "Post Mile": 11, "Length": 9, "Description": 26}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26
                             else "A" + chr(64 + i - 26)].width = widths.get(name, 8)
    hdc.apply_header_tooltips(ws)

    prov = wb.create_sheet(PROVENANCE_SHEET)
    prov.sheet_properties.tabColor = "1F3864"
    prov.append(["Report column", "CA HIGHWAYS column(s)", "How", "Tier"])
    for c in next(prov.iter_rows(min_row=1, max_row=1)):
        c.font, c.fill = head_font, head_fill
    for row in _provenance_rows():
        prov.append(list(row))
    for col, w in (("A", 16), ("B", 46), ("C", 34), ("D", 14)):
        prov.column_dimensions[col].width = w

    mk = wb.create_sheet(MARKER_SHEET)
    mk.sheet_properties.tabColor = "1F3864"
    lines = [
        ("Build version", BUILD_VERSION),
        ("As-of date", facts.get("asof", "(unknown)")),
        ("Source workbook", str(source_path)),
        ("Source build version", facts.get("source_build_version", "(unknown)")),
        ("Layer library", facts.get("layer_library", "(unknown)")),
        ("Layer drop exported", facts.get("drop_exported") or "(unknown)"),
        ("Layer drop fingerprint", facts.get("drop_fingerprint") or "(unknown)"),
        ("CA HIGHWAYS rows read", stats["source_rows"]),
        ("Report records written", stats["rows"]),
        ("Rows merged away", stats["merged_away"]),
        ("Routes", stats["routes"]),
        ("Merge rule", "adjacent spans identical across every printed column "
                       "and touching exactly are ONE record; the printed Length "
                       "is the merged span's end PM - begin PM"),
        ("Columns with no source", ", ".join(CONTEXT_COLUMNS) or "(none)"),
    ]
    # Carried from the source build so the comparison can read the condition
    # off the side it actually loads, without opening the CA HIGHWAYS workbook
    # a second time (it may not even still be on disk).
    if _skipped_spans(facts):
        lines += [
            ("Skipped source spans", facts.get("skipped_source_spans", "")),
            ("Marked anchor cells", facts.get("marked_anchor_cells", "")),
            ("Unavailable marker", facts.get("unavailable_marker", "")),
            ("Skipped source reason", facts.get("skipped_source_reason", "")),
        ]
    for k, v in lines:
        mk.append([k, v])
    mk.column_dimensions["A"].width = 24
    mk.column_dimensions["B"].width = 70
    for row in mk.iter_rows(min_col=2, max_col=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    hdc.write_legend_sheet(wb)
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# public entry point
# --------------------------------------------------------------------------- #
SOURCE_TABLE = "highway_detail_source_table.xlsx"


def consolidate(events=None, confirm_overwrite=None, day=None, *,
                built_path=None, out_path=None, asof=None, lib_root=None):
    """Build the Highway Detail report FROM THE ARCGIS LAYERS.

    The report is built in two measured passes over the layer library, not by
    reading the Clean Road output:

      1. the shared span engine over THIS REPORT'S OWN layers
         (`SEGMENT_TAGS`) with the report's own block-eff-date rule
         (`primary_eff`), producing its CA HIGHWAYS source table;
      2. the projection onto the report's 34 printed columns.

    It reads the Clean Road build ONLY when `built_path` is given explicitly
    (verification, and the golden check's synthetic fixtures). That separation is
    the point: the Clean Road build reproduces the TSN table, so it segments on
    TSN's 74 columns and dates its blocks the way TSN's own table was measured.
    A REPORT wants its own columns' boundaries and its own measured rule, and
    tying the two made the report inherit ~6,500 splits it does not print plus a
    block-date rule that scored 56-60% against the real export (DA1).

    `asof` is the reconstruction date — set it to the export day being compared
    against, or the comparison measures network change instead of correctness.
    `out_path` defaults to `output/arcgis_reports/`; the destination is app-owned
    and a rebuild replaces it by design."""
    del day                       # ConsolidateWorker passes it; no run days here
    events = events or Events()
    try:
        if built_path is None:
            built_path = _build_source_table(events, out_path, asof, lib_root)
            if isinstance(built_path, ConsolidateResult):
                return built_path      # the build failed / was cancelled
        return _consolidate(events, confirm_overwrite, built_path, out_path)
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))


def _build_source_table(events, out_path, asof, lib_root):
    """Pass 1: this report's own CA HIGHWAYS table, from the layers. Returns its
    path, or the failed ConsolidateResult to hand straight back.

    Kept on disk beside the report (not a temp) because it IS the report's
    source of record: the Provenance and marker sheets point at it, and a
    question about a printed cell is answered by the row it came from."""
    dest = (Path(out_path).parent if out_path else OUT_DIR) / SOURCE_TABLE
    dest.parent.mkdir(parents=True, exist_ok=True)
    res = cch.consolidate(events=events, confirm_overwrite=lambda p: True,
                          asof=asof, lib_root=lib_root, out_path=dest,
                          span_tags=SEGMENT_TAGS, primary_eff=True)
    if res.status != "ok":
        return res
    return dest


def _consolidate(events, confirm_overwrite, built_path, out_path):
    source = Path(built_path) if built_path else cch.OUT_PATH
    out_path = Path(out_path) if out_path else OUT_PATH

    if out_path.exists() and confirm_overwrite is not None \
            and not confirm_overwrite(out_path):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.",
                                 completion=outcome.CANCELLED)

    facts = built_facts(source) if source.is_file() else {}
    asof = facts.get("asof") or "(unknown)"
    events.on_log("=" * 60)
    events.on_log(f"{REPORT_NAME} — projected from the Clean Road build "
                  f"(as of {asof})")
    events.on_log("=" * 60)

    rows, stats = project_rows(_require_built(source), events)
    if events.is_cancelled():
        return ConsolidateResult(status="cancelled", message="Cancelled.",
                                 completion=outcome.CANCELLED)
    if not rows:
        return ConsolidateResult(
            status="error",
            message="The Clean Road Highway build produced no rows to project — "
                    "rebuild it and try again.")

    _write_workbook(out_path, rows, stats, facts, source)
    # A projection is exactly as assertable as the build under it. When the CA
    # HIGHWAYS build could not place a span, its marker token is sitting in
    # this report's cells too, so this result must not read COMPLETE over it
    # (the completion is what the comparison and the sidecar go by).
    skipped = _skipped_spans(facts)
    marked = facts.get("marked_anchor_cells", "")
    result = ConsolidateResult(
        status="ok",
        message=(f"Built {stats['rows']:,} Highway Detail records "
                 f"({stats['routes']} routes) from the ArcGIS layers as of "
                 f"{asof}. {stats['merged_away']:,} CA HIGHWAYS row(s) merged "
                 f"into a neighbouring record (they split only on columns the "
                 f"report does not print)."
                 + (f" The CA HIGHWAYS build could not place {skipped:,} "
                    f"source span(s), so {marked} cell(s) here carry "
                    f"'{facts.get('unavailable_marker', '')}' and assert "
                    f"nothing." if skipped else "")),
        output_path=str(out_path),
        summary_lines=[f"{REPORT_NAME}: {stats['rows']:,} records, "
                       f"{stats['routes']} routes -> {out_path.name}"],
        skipped_inputs=skipped,
        completion=outcome.PARTIAL if skipped else outcome.COMPLETE)
    if not consolidation_meta.write_outcome(
            out_path, result,
            extra={SIDECAR_KEY: {
                "report": "highway_detail",
                "build_version": BUILD_VERSION,
                "asof": facts.get("asof", ""),
                "layer_drop": {"fingerprint": facts.get("drop_fingerprint") or None,
                               "exported": facts.get("drop_exported") or None},
                "source_workbook": str(source),
                "source_build_version": facts.get("source_build_version", ""),
                "context_columns": list(CONTEXT_COLUMNS),
                "skipped_source_spans": skipped,
                "marked_anchor_cells": marked,
                "unavailable_marker": facts.get("unavailable_marker", ""),
                **stats,
            }}):
        return ConsolidateResult(
            status="error",
            message="The build finished but its outcome sidecar could not be "
                    "published — rebuild before comparing.")
    result.sidecar_published = True
    events.on_log(result.message)
    return result
