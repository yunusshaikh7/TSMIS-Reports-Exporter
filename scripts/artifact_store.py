"""Transactional artifact lifecycle (P2) — one leaf module for three concerns:

  * **Atomic single-file write** (`atomic_save`, `atomic_save_if`, `commit_workbook`) —
    write to a temp sibling, validate, then ``os.replace`` it onto the final path, so an
    interrupted / failed / locked write NEVER truncates the prior good artifact (F9). The
    wrapper is handed a TEMP path and finalizes it; ``compare_core`` writes to that
    path and now rechecks the same optional ownership guard at its exact save boundary.
    Workbook formulas/layout remain regression-locked. The wrapper rewrites any leaked
    temp name back out of the returned result. `confirm_late_overwrite`
    + `atomic_save_if` add the P12 pre-commit re-check that closes the consolidate
    confirm-then-appears TOCTOU window (the truncation half is already covered by F9).

  * **Journaled store promotion + startup recovery** (`promote_store`,
    `recover_promotions`) — the Export-Everything store swap (`live` <- `staging`) is
    journaled so a crash BETWEEN the two renames is repaired on the next launch from the
    retained backup: there is never a window with zero copies (F2). Recovery is
    idempotent and runs from `updater.cleanup_leftovers`.

  * **Input fingerprint** (`fingerprint`, `consolidated_fresh`) — freshness tracks input
    IDENTITY, not newest-mtime: a hash of the store's sorted ``(name, size, mtime_ns)``
    (R1-R03). A DELETED (non-newest) route changes the fingerprint though it leaves the
    newest-mtime untouched, so a stale consolidated no longer reads as fresh (F5).

Console-free (the core contract): reports via ``log`` + return values / raises, never
print/input/exit. Module-level imports are stdlib plus the dependency-light comparison
contracts and ``events`` result shape; ``openpyxl`` is imported LAZILY inside
``_openable_xlsx`` (workbook validation opens the produced file to reject a
malformed/unreadable XLSX before committing it).
"""
import hashlib
import json
import logging
import os
import re
import stat as statmod
import time
import uuid
import zipfile
from dataclasses import dataclass, replace
from pathlib import Path

from comparison_contract import ArtifactGeneration, AttemptState, ComparisonOutcome
from events import ConsolidateResult

log = logging.getLogger("tsmis.artifact_store")

# Sidecar written beside a consolidated workbook recording its inputs' fingerprint.
_FP_SUFFIX = ".fingerprint.json"
# v2 (CMP-AUD-080): the folder fingerprint hashes file CONTENT, not
# (size, mtime_ns). Bumping the schema is what migrates every v1 record to stale
# exactly once.
_FP_SCHEMA = 2
# fingerprint() sentinel: the folder (or a file in it) could not be read -> the caller
# must treat freshness CONSERVATIVELY (rebuild), never as a silent match.
_UNREADABLE = "unreadable"
# Names excluded from a store fingerprint: Excel lock files, our own sidecars, and the
# in-flight temp / staging siblings this module itself creates.
_FP_EXCLUDED_SUFFIXES = (_FP_SUFFIX, ".outcome.json", ".provenance.json",
                         ".staging")
_COMPARISON_PUBLICATION_LOCK_NAME = ".tsmis-comparison-publication.lock"
# Both chunk-name shapes (CMP-AUD-242): the short 16-hex form current builds
# write, and the legacy full-hex form earlier builds left behind. Keep in sync
# with consolidation_meta._PAYLOAD_BASENAME_RE.
_COMPARISON_PAYLOAD_RE = re.compile(
    r"^\.cmpv3-(?:[0-9a-f]{16}|[0-9a-f]{64})-[0-9]{6}-(?:[0-9a-f]{16}|[0-9a-f]{64})"
    r"(?:-f-(?:0[0-7]|[0-9a-f]{64}-[0-9a-f]{16}))?"
    r"\.comparison-payload\.zlib$")
# The per-route export artifacts a store legitimately contains (P2-R01: staging must hold
# a real report file, not just any regular file). XLSX for every report; PDF for the
# Ramp Summary and Highway Log PDF stores.
_REPORT_SUFFIXES = (".xlsx", ".pdf")
_PRODUCER_TEMP_INFIX = ".tmp-"
_PRODUCER_TEMP_TOKEN_HEX_CHARS = 12
_VALUES_TWIN_LABEL = " (values)"


def _new_token():
    """A short unique token for temp / backup names (per write, per promotion)."""
    return uuid.uuid4().hex[:_PRODUCER_TEMP_TOKEN_HEX_CHARS]


def _silent_unlink(path):
    """Best-effort unlink; never raises. True iff the file is gone afterwards."""
    if path is None:
        return True
    try:
        Path(path).unlink()
        return True
    except FileNotFoundError:
        return True
    except OSError:
        return False


def _values_twin(path):
    """The values-flavor sibling compare_core derives from a picked name:
    ``<stem> (values)<suffix>`` (mirrors compare_core.run_compare)."""
    path = Path(path)
    return path.with_name(f"{path.stem}{_VALUES_TWIN_LABEL}{path.suffix}")


def _producer_temp(path, token):
    """Derive the exact producer-temp sibling used by every workbook commit."""
    path = Path(path)
    return path.with_name(
        f"{path.stem}{_PRODUCER_TEMP_INFIX}{token}{path.suffix}")


def _comparison_path_limit_error(final, twin):
    """Return an actionable pre-mutation error, or ``None`` when names fit.

    The check constructs every mandatory basename through the same helpers used
    by the producer and metadata publisher.  A fixed all-zero token is safe here
    because `_new_token` is fixed-width lowercase hex; content never changes its
    UTF-16 length.
    """
    import consolidation_meta

    final = Path(final)
    finals = (final, _values_twin(final)) if twin else (final,)
    token = "0" * _PRODUCER_TEMP_TOKEN_HEX_CHARS
    try:
        for workbook in finals:
            consolidation_meta.validate_comparison_metadata_paths(workbook)
        producer_primary = _producer_temp(final, token)
        producer_paths = ((producer_primary, _values_twin(producer_primary))
                          if twin else (producer_primary,))
        for producer_path in producer_paths:
            consolidation_meta._require_windows_component(
                producer_path.name, "producer temporary workbook")
    except ValueError as e:
        return (
            f"Cannot create this comparison: {e}. No comparison producer was "
            "started and existing files were kept.")
    return None


def _resolved_identity(path):
    """Best-effort canonical text identity for an output/source path.

    ``strict=False`` deliberately handles a destination that does not exist yet.
    The app is Windows-only, so case-folding is conservative and closes aliases
    whose spelling differs only by case even when this helper is exercised on a
    case-sensitive development filesystem.
    """
    path = Path(path)
    try:
        resolved = path.resolve(strict=False)
    except (OSError, RuntimeError):       # silent-ok: abspath is the conservative canonical fallback
        resolved = Path(os.path.abspath(os.fspath(path)))
    return os.path.normpath(os.fspath(resolved)).casefold()


@dataclass(frozen=True)
class _CapturedSource:
    """The filesystem object selected as a comparison source at entry."""
    path: Path
    file_id: object
    is_dir: bool


def _stat_identity(path):
    """Return ``((device, inode), is_directory)`` or ``(None, False)`` if absent.

    Windows' ``st_ino`` is the volume file index, so this remains stable across
    a rename and is the same identity primitive used by ``os.path.samefile``.
    Any other stat failure is uncertainty, not evidence that a write is safe.
    """
    path = Path(path)
    try:
        st = path.stat()
    except (FileNotFoundError, NotADirectoryError):  # silent-ok: absence is an explicit identity state
        return None, False
    except OSError as e:
        raise ValueError(
            f"Cannot verify comparison source identity for {path}: "
            f"{type(e).__name__}. Choose an accessible local source.") from e
    return (st.st_dev, st.st_ino), statmod.S_ISDIR(st.st_mode)


def capture_source_identities(source_paths):
    """Capture stable file IDs for every selected comparison source.

    Missing paths are retained as a ``None`` identity so the adapter can return
    its existing source-missing message; if such a path appears during a
    successful producer run, the later current-source check fails closed.
    """
    captured = []
    for raw in (source_paths or ()):
        if raw is None:
            continue
        path = Path(raw)
        file_id, is_dir = _stat_identity(path)
        captured.append(_CapturedSource(path, file_id, is_dir))
    return tuple(captured)


def canonical_path_identities(paths):
    """Case-folded resolved path set for a discovery-membership tripwire."""
    return frozenset(_resolved_identity(path) for path in (paths or ()))


def _destination_identity_candidates(path):
    """Yield a destination and its lexical/resolved ancestors, without repeats."""
    path = Path(path)
    try:
        resolved = path.resolve(strict=False)
    except (OSError, RuntimeError) as e:
        raise ValueError(
            f"Cannot verify comparison output identity for {path}: "
            f"{type(e).__name__}. Choose an accessible local destination.") from e
    seen = set()
    for base in (path, resolved):
        for candidate in (base, *base.parents):
            key = _resolved_identity(candidate)
            if key not in seen:
                seen.add(key)
                yield candidate


def _destination_aliases_captured(dest, captured):
    """True if ``dest`` is, or is below, the captured source object."""
    candidates = ((Path(dest),) if not captured.is_dir
                  else _destination_identity_candidates(dest))
    for candidate in candidates:
        file_id, _is_dir = _stat_identity(candidate)
        if file_id is not None and file_id == captured.file_id:
            return True
    return False


def _paths_alias(left, right):
    """True when two paths name (or resolve to) the same filesystem object.

    Canonical text covers relative paths, ``..``, case-only spellings, and
    symlink/junction resolution even when the output is not present yet.
    ``samefile`` additionally catches hardlinks once both objects exist.
    """
    if _resolved_identity(left) == _resolved_identity(right):
        return True
    try:
        return os.path.samefile(left, right)
    except (FileNotFoundError, NotADirectoryError, ValueError):  # silent-ok: a missing side cannot share a live file ID
        return False
    except OSError:                       # silent-ok: identity uncertainty is handled by rejecting the write
        # An existing object's identity that the OS refuses to reveal is not
        # evidence of safety.  Reject conservatively instead of overwriting it.
        return True


def _identity_within(path_identity, root_identity):
    """True when canonical identity text is equal to or below a directory."""
    try:
        return os.path.commonpath((path_identity, root_identity)) == root_identity
    except ValueError:                    # silent-ok: different drives cannot be path-contained
        return False


def ensure_outputs_do_not_alias_sources(destinations, source_paths, *,
                                         directory_destinations=(),
                                         captured_sources=(),
                                         require_sources_current=False):
    """Reject a write destination that could overwrite a comparison source.

    Besides exact/same-file aliases, a destination *inside* a source directory
    is rejected: folder comparisons hand their selected roots to this boundary,
    and allowing a comparison workbook there could replace a discovered input.
    For directory destinations (the visual-evidence image set), the inverse is
    rejected too because swapping that directory would delete a nested source.

    The check is intentionally safe to call twice: once before expensive work and
    again immediately before commit to catch a destination that appeared as a
    hardlink while the producer ran.
    """
    destinations = tuple(Path(p) for p in destinations if p is not None)
    sources = tuple(Path(p) for p in (source_paths or ()) if p is not None)
    directory_ids = {_resolved_identity(p) for p in directory_destinations
                     if p is not None}
    for dest in destinations:
        dest_id = _resolved_identity(dest)
        for source in sources:
            source_id = _resolved_identity(source)
            aliases = _paths_alias(dest, source)
            if not aliases:
                try:
                    source_is_dir = source.is_dir()
                except OSError:           # silent-ok: an unreadable source is not assumed to be a directory
                    source_is_dir = False
                if source_is_dir:
                    aliases = _identity_within(dest_id, source_id)
            if not aliases and dest_id in directory_ids:
                aliases = _identity_within(source_id, dest_id)
            if aliases:
                raise ValueError(
                    f"Refusing to write {dest}: that output aliases comparison "
                    f"source {source}. Choose an output outside the selected inputs.")
        for captured in captured_sources:
            if (captured.file_id is not None
                    and _destination_aliases_captured(dest, captured)):
                raise ValueError(
                    f"Refusing to write {dest}: that output aliases the originally "
                    f"selected comparison source {captured.path}, even though the "
                    "source may have moved. Re-run after the source paths are stable.")

    if require_sources_current:
        for captured in captured_sources:
            current_id, _is_dir = _stat_identity(captured.path)
            if current_id != captured.file_id:
                raise ValueError(
                    f"Refusing to finalize the comparison: source {captured.path} "
                    "changed identity or location while the comparison was running. "
                    "The prior output was kept; re-run with stable source files.")


def comparison_output_paths(out_path, mode):
    """Return every final workbook path a comparison mode can publish.

    ``run_compare`` derives the values sibling only for ``mode='both'``. Keep
    that naming in one dependency-light boundary so public file/folder drivers
    can guard the unselected sibling before loading or writing source data.
    Invalid modes retain the primary path only; ``run_compare`` still owns the
    existing user-facing invalid-mode result.
    """
    out = Path(out_path)
    return (out, _values_twin(out)) if mode == "both" else (out,)


# --------------------------------------------------------------------------- #
# atomic single-file write (F9)
# --------------------------------------------------------------------------- #
def atomic_save(workbook, out_path):
    """Save an openpyxl `workbook` to a temp sibling of `out_path`, then atomically
    ``os.replace`` it onto `out_path`. An interrupted or failed write (disk error, or
    the DESTINATION open in Excel -> ``PermissionError`` on ``os.replace``) leaves the
    prior `out_path` UNTOUCHED (F9: never truncate a good prior artifact). The original
    exception propagates to the caller's existing handler AFTER the temp is removed, so
    a consolidator's ``except PermissionError`` keeps reporting "file open in Excel"."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = _producer_temp(out_path, _new_token())
    try:
        workbook.save(tmp)
        os.replace(tmp, out_path)
    except BaseException:
        _silent_unlink(tmp)
        raise


def atomic_save_if(workbook, out_path, proceed):
    """Like `atomic_save`, but the final ``os.replace`` is GATED on `proceed()` —
    a 0-arg callback evaluated AFTER the workbook is fully serialized to the temp
    sibling and JUST BEFORE the replace. If `proceed()` is falsy the temp is removed
    and `out_path` is left untouched; returns True iff the replace happened.

    This lets a producer run its final TOCTOU re-check (confirm_late_overwrite) at
    the NARROWEST possible point — the workbook is already written to the temp, so
    there is no half-streamed ``write_only`` workbook to abandon (which would leave
    an open temp + a dangling row generator). Same F9 guarantee as `atomic_save`: a
    prior good `out_path` is never truncated; the temp is removed on any exit."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = _producer_temp(out_path, _new_token())
    try:
        workbook.save(tmp)
        if not proceed():
            _silent_unlink(tmp)
            return False
        os.replace(tmp, out_path)
        return True
    except BaseException:
        _silent_unlink(tmp)
        raise


def _openable_xlsx(path, expect_sheet=None):
    """A produced workbook is committable iff openpyxl can actually OPEN it — so a corrupt
    ZIP or a malformed workbook part (e.g. ``xl/workbook.xml`` = ``b"not xml"``) is
    REJECTED, not just name-checked (P2-B05) — AND it has >=1 sheet (and `expect_sheet`,
    when given, is present: the per-comparison sheet contract). ``read_only`` load is lazy
    (it parses ``workbook.xml`` + the sheet list, not the cells), so it stays cheap even
    for a big live-formulas workbook. Never raises — any open/parse failure ⇒ invalid.

    The workbook is opened on a caller-owned file object closed by ``with`` (P2-R04): if
    ``load_workbook`` raises on a malformed part it would otherwise leave openpyxl's ZIP
    handle open, locking the temp on Windows so cleanup could not remove it. Closing OUR
    handle on every exit releases the OS lock regardless of where the parse failed."""
    try:
        p = Path(path)
        if not p.is_file() or p.stat().st_size == 0 or not zipfile.is_zipfile(p):
            return False
        from openpyxl import load_workbook
        with open(p, "rb") as fh:                    # our handle -> closed on ANY exit (incl. raise)
            wb = load_workbook(fh, read_only=True)
            try:
                names = list(wb.sheetnames)
            finally:
                wb.close()
        if not names:
            return False
        return expect_sheet is None or expect_sheet in names
    except Exception:                                # noqa: BLE001 — open/parse failure => invalid
        return False


def _commit_one(tmp, final, validate, proceed=None, discard=None,
                temp_current=None, final_current=None):
    """Validate `tmp` then atomically replace it onto `final`. True iff committed.
    A missing/empty/invalid temp or a failed replace leaves `final` untouched, removes
    the temp, and returns False (never raises). ``proceed`` is an optional last-moment
    safety predicate evaluated after validation and immediately before ``os.replace``.
    ``temp_current`` binds the exclusively-created producer file across validation;
    ``final_current`` is the target-aware ownership check at publication."""
    tmp, final = Path(tmp), Path(final)
    discard = discard or _silent_unlink
    if temp_current is not None and not temp_current():
        discard(tmp)
        return False
    if not validate(tmp):
        if not discard(tmp):                         # P2-R04: verify the rejected temp is gone
            log.warning("artifact commit: rejected temp %s could not be removed (locked?)",
                        tmp.name)
        return False
    try:
        # Validation opens the pathname and may be slow.  A same-name ordinary
        # replacement must never become the object we publish merely because it
        # is also a valid workbook.
        if temp_current is not None and not temp_current():
            discard(tmp)
            return False
        if final_current is not None and not final_current():
            discard(tmp)
            return False
        if proceed is not None and not proceed():
            discard(tmp)
            return False
        if temp_current is not None and not temp_current():
            discard(tmp)
            return False
        if final_current is not None and not final_current():
            discard(tmp)
            return False
        os.replace(tmp, final)
        return True
    except OSError as e:
        log.warning("artifact commit: could not finalize %s (%s: %s); prior kept",
                    final.name, type(e).__name__, e)
        discard(tmp)
        return False


def _rewrite_paths(result, mapping):
    """Rewrite leaked temp paths in a ConsolidateResult's ``output_path``, ``message`` AND
    ``summary_lines`` back to their final names (so a ``.tmp-<token>`` name NEVER surfaces
    — including in an ERROR result's message, which compare_core builds from the path it
    was handed; P2-R02). The wrapper handed the producer temp paths; the user must see the
    real destinations. Applied to EVERY returned status (ok / error / cancelled)."""
    def sub(text):
        for tmp, final in mapping.items():
            text = text.replace(tmp, final)
        return text
    if getattr(result, "output_path", ""):
        result.output_path = sub(result.output_path)
    if getattr(result, "message", ""):
        result.message = sub(result.message)
    if getattr(result, "summary_lines", None):
        result.summary_lines = [sub(s) for s in result.summary_lines]
    return result


def confirm_late_overwrite(dest, existed_at_confirm, confirm):
    """TOCTOU re-check (P12): close the confirm-then-appears window.

    A destination that did NOT exist when the user was first asked about
    overwriting (so they were never prompted for it) can APPEAR while the producer
    runs — parsing a folder of PDFs or building a large workbook takes seconds. The
    original code then ran ``os.replace`` and silently clobbered the file the user
    never agreed to overwrite. Re-checked here, just before the commit: if it
    appeared, ask now. Returns True to proceed (it didn't appear, or the user
    re-confirmed) or False to abort (the user declined the newly-appeared file).

    This NARROWS the window from the whole producer runtime to the microseconds
    between this check and ``os.replace``; it cannot ATOMICALLY eliminate it (there
    is no "replace only if still absent" for our deliberate overwrite-after-confirm
    case). The TRUNCATION half of the original TOCTOU is already closed — the
    producer writes a temp sibling and we atomically ``os.replace`` it, so a prior
    good artifact is never truncated (F9/P2). `confirm(dest)->bool` is the SAME
    callback used for the first prompt (default: overwrite freely)."""
    if existed_at_confirm:
        return True                          # already confirmed at the first ask
    if not Path(dest).exists():
        return True                          # still absent — nothing appeared
    return bool(confirm(dest))               # appeared during produce — ask now


def _artifact_stat_signature(st):
    """Mutation-sensitive fields bound to a committed generation member."""
    return (st.st_dev, st.st_ino, st.st_size, st.st_mtime_ns)


def _committed_artifact_member(path, *, flavor, commit_role, current_guard):
    """Hash one committed ordinary file through an identity-bound read-only fd.

    A pathname can be replaced or its bytes can change while a digest is being
    calculated.  Capture both path and descriptor identity, compare size/mtime
    before and after the read, then re-run the transaction's exact target guard.
    No generation claim is returned if any boundary changes.
    """
    path = Path(path)
    identity = _plain_entry_identity(path, directory=False)
    if identity is None:
        raise ValueError(f"committed {flavor} artifact is not an ordinary file")
    try:
        before = path.lstat()
    except OSError as e:
        raise ValueError(
            f"could not stat committed {flavor} artifact ({type(e).__name__})") from e
    before_sig = _artifact_stat_signature(before)
    if before_sig[:2] != identity:
        raise ValueError(f"committed {flavor} artifact changed before hashing")

    flags = os.O_RDONLY | getattr(os, "O_BINARY", 0)
    flags |= getattr(os, "O_NOINHERIT", 0) | getattr(os, "O_NOFOLLOW", 0)
    fd = None
    try:
        fd = os.open(path, flags)
        if _plain_fd_identity(fd) != identity:
            raise ValueError(f"committed {flavor} artifact changed while opening")
        opened = os.fstat(fd)
        if _artifact_stat_signature(opened) != before_sig:
            raise ValueError(f"committed {flavor} artifact changed before reading")
        digest = hashlib.sha256()
        while True:
            chunk = os.read(fd, 1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
        after_fd = os.fstat(fd)
        if _artifact_stat_signature(after_fd) != before_sig:
            raise ValueError(f"committed {flavor} artifact changed while hashing")
    except OSError as e:
        raise ValueError(
            f"could not hash committed {flavor} artifact ({type(e).__name__})") from e
    finally:
        if fd is not None:
            os.close(fd)

    try:
        after_path = path.lstat()
    except OSError as e:
        raise ValueError(
            f"could not re-stat committed {flavor} artifact ({type(e).__name__})") from e
    if (_plain_entry_identity(path, directory=False) != identity
            or _artifact_stat_signature(after_path) != before_sig):
        raise ValueError(f"committed {flavor} artifact changed after hashing")
    if not current_guard(path):
        raise ValueError(
            f"destination ownership changed after hashing the {flavor} artifact")
    try:
        final_stat = path.lstat()
    except OSError as e:
        raise ValueError(
            f"could not verify committed {flavor} artifact ({type(e).__name__})") from e
    if (_plain_entry_identity(path, directory=False) != identity
            or _artifact_stat_signature(final_stat) != before_sig):
        raise ValueError(f"committed {flavor} artifact changed after its final guard")

    relative_path = path.name
    if (not relative_path or relative_path in (".", "..")
            or "/" in relative_path or "\\" in relative_path
            or ":" in relative_path
            or Path(relative_path).name != relative_path):
        raise ValueError("committed artifact basename is not a safe relative path")
    return {
        "flavor": flavor,
        "path": str(path),
        "relative_path": relative_path,
        "canonical_path_at_write": _resolved_identity(path),
        "commit_role": commit_role,
        "sha256": digest.hexdigest(),
        "size": before.st_size,
        "mtime_ns": before.st_mtime_ns,
    }


def _attach_artifact_generation(result, members, requested_mode):
    """Attach one exact committed-member generation to a typed result."""
    typed = getattr(result, "comparison_outcome", None)
    if not isinstance(typed, ComparisonOutcome):
        return result
    members = tuple(dict(member) for member in members)
    generation_id = str(uuid.uuid4())
    result.artifact_generation = ArtifactGeneration(
        generation_id=generation_id,
        members=members,
        content_digests={member["flavor"]: member["sha256"] for member in members},
        completion=typed.completion,
        producer_versions={},
        publication_state="committed",
        requested_mode=requested_mode,
    )
    result.attempt_state = AttemptState(
        state="succeeded", generation_id=generation_id)
    return result


def _publish_artifact_generation(result, commit_guard=None):
    """Publish typed comparison metadata beside every committed member.

    Publication failure changes the attempt/publication axes, not the immutable
    comparison counts and coverage represented by the committed workbook.
    """
    generation = getattr(result, "artifact_generation", None)
    if not isinstance(generation, ArtifactGeneration):
        return result
    try:
        import consolidation_meta
        published = consolidation_meta.write_comparison_outcomes(
            result, commit_guard=commit_guard)
    except Exception:  # noqa: BLE001 - turn publication defects into a visible fail-closed result
        log.exception("comparison generation sidecar publication crashed")
        published = False
    if published:
        return result

    message = (
        "The comparison workbook was created, but its generation metadata could "
        "not be safely published for every output member. The files remain marked "
        "untrusted; close any open copies and run the comparison again. If this "
        "repeats for every comparison, the output folder path is probably too "
        "long for Windows — move the app and its output to a shorter folder path "
        "(the log names the exact file and its length).")
    result.artifact_generation = replace(generation, publication_state="partial")
    result.attempt_state = AttemptState(
        state="failed", message=message, generation_id=generation.generation_id)
    if getattr(result, "status", None) == "ok":
        result.status = "error"
        result.message = message
    elif message not in str(getattr(result, "message", "") or ""):
        prior = str(getattr(result, "message", "") or "").strip()
        result.message = f"{prior}\n\n{message}" if prior else message
    return result


# --------------------------------------------------------------------------- #
# CMP-AUD-115 — the versioned COMPARISON-ARTIFACT SCHEMA gate.
#
# The transactional commit used to require only that openpyxl could open the
# workbook and that a sheet named `Comparison` existed, so a header-only or
# label-less Comparison sheet published with status=ok/verdict=match.
#
# The schema below is deliberately the SAME contract the Matrix count reader
# already enforces (unique `Status`/`Diffs` labels, a valid status on every data
# row, an integer `Diffs` on a matched row and none on a one-sided one), so the
# gate's rejection domain is a SUBSET of the already-unreadable domain: a
# workbook this refuses is one the Matrix would have read as `(None, None)`
# anyway. It therefore cannot block a report that would otherwise have worked —
# it converts a silently unreadable artifact into a loud, kept-last-good commit
# failure. It applies ONLY to a typed comparison's VALUES artifact: the
# live-formulas twin holds formulas rather than cached values by construction,
# and a consolidation carries no Comparison sheet at all.
# --------------------------------------------------------------------------- #
COMPARISON_ARTIFACT_SCHEMA = 1


def comparison_counts(values_path):
    """``(diff_cells, one_sided, data_rows)`` from a VALUES comparison workbook,
    or ``(None, None, None)``.

    ``Status`` and ``Diffs`` are the structured truth columns emitted by every
    current comparison layout, located by UNIQUE EXACT LABEL. Visible field text
    is never inspected: the spaced not-equal glyph is legitimate source content
    and cannot encode state. A missing, duplicate, or malformed count contract
    returns the unknown triple instead of guessing a layout or silently
    certifying zero differences."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(values_path, read_only=True, data_only=True)
    except Exception as e:                       # noqa: BLE001 (best-effort read)
        log.debug("comparison_counts: can't open %s (%s: %s)", values_path,
                  type(e).__name__, e)
        return (None, None, None)
    try:
        ws = wb["Comparison"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None) or ()     # row 1 (header); data follows

        def _col_of(label):                      # unique 1-based exact label
            matches = [i + 1 for i, value in enumerate(header)
                       if value == label]
            return matches[0] if len(matches) == 1 else None

        status_col = _col_of("Status")
        diffs_col = _col_of("Diffs")
        if status_col is None or diffs_col is None:
            log.debug("comparison_counts: Comparison lacks unique Status/Diffs labels")
            return (None, None, None)
        one_sided = diff_cells = data_rows = 0
        for row in rows_iter:                     # data rows (header consumed above)
            if row is None or all(v is None for v in row):
                continue
            data_rows += 1
            status = row[status_col - 1] if len(row) >= status_col else None
            diffs = row[diffs_col - 1] if len(row) >= diffs_col else None
            if status == "Both":
                if (isinstance(diffs, bool) or not isinstance(diffs, (int, float))
                        or not float(diffs).is_integer() or diffs < 0):
                    log.debug("comparison_counts: matched row has invalid Diffs value %r", diffs)
                    return (None, None, None)
                diff_cells += int(diffs)
            elif isinstance(status, str) and status:
                if diffs not in (None, ""):
                    log.debug("comparison_counts: one-sided row unexpectedly carries Diffs %r", diffs)
                    return (None, None, None)
                one_sided += 1
            else:
                log.debug("comparison_counts: row has invalid Status value %r", status)
                return (None, None, None)
        return (diff_cells, one_sided, data_rows)
    except Exception as e:                       # noqa: BLE001
        log.debug("comparison_counts: can't read %s (%s: %s)", values_path,
                  type(e).__name__, e)
        return (None, None, None)
    finally:
        wb.close()


def _typed_row_claim(typed_outcome):
    """How many rows the producer's typed outcome says the comparison covered."""
    counts = getattr(typed_outcome, "counts", None)
    if counts is None or not getattr(counts, "known", False):
        return 0
    try:
        return (int(counts.paired_rows) + int(counts.side_a_only_rows)
                + int(counts.side_b_only_rows))
    except (TypeError, ValueError):  # silent-ok: an unusable claim simply asserts nothing here
        return 0


def comparison_artifact_problem(values_path, typed_outcome=None):
    """None when `values_path` satisfies COMPARISON_ARTIFACT_SCHEMA, else a
    one-line reason for the commit refusal."""
    diff_cells, _one_sided, data_rows = comparison_counts(values_path)
    if diff_cells is None:
        return ("its Comparison sheet does not carry uniquely labelled Status and "
                "Diffs columns with a valid status on every row")
    if data_rows == 0 and _typed_row_claim(typed_outcome) > 0:
        return ("its Comparison sheet has no rows although the comparison "
                "reported paired or one-sided rows")
    return None


def commit_workbook(final, produce_fn, *, twin=False, expect_sheet=None, validate=None,
                     confirm_overwrite=None, source_paths=(), captured_sources=None,
                     commit_guard=None, requested_mode=None):
    """Run `produce_fn(temp_path)` — the EXISTING writer (compare_core via an adapter),
    pointed at a temp sibling of `final` — then validate and atomically commit it onto
    `final`. The producer writes only to the temp path it is handed; this wrapper
    finalizes and rewrites the temp name out of the result (F9).
    Every returned result (ok / error / cancelled) is path-sanitized, so a ``.tmp-<token>``
    name never reaches the user (P2-R02).

    Validation OPENS the produced workbook (not a name check) and, with `expect_sheet`,
    requires that sheet — a malformed/corrupt output is rejected and the prior `final` is
    kept (P2-B05). `confirm_overwrite(path)->bool` is checked against the FINAL
    destination(s) BEFORE producing; a decline returns a ``cancelled`` result.
    Every final destination is also checked against ``source_paths`` before
    production and again before commit; canonical-path and same-file identity
    reject direct, symlink/junction, case-only, and existing hardlink aliases.

    `twin=True` (a ``mode="both"`` comparator): the producer writes BOTH the formulas
    workbook (the temp primary) and its ``(values)`` sibling. Per the multi-file policy
    (Q5) the **values** workbook is the single transactional artifact — committed FIRST;
    the **formulas** sibling is best-effort, committed second. If the formulas commit
    fails the result is rewritten to be TRUTHFUL: ``output_path`` points at the committed
    values workbook and the formulas line becomes a not-refreshed warning (P2-R03). A
    failure to commit the transactional (values, or the lone file) artifact leaves the
    prior `final` untouched and returns an error result.

    ``commit_guard`` is target-aware when supplied: ``guard(path, **binding)``.
    Every parent/final/temp boundary is checked. Callbacks that accept ``path``
    but not the optional binding keywords remain compatible (the transaction's
    own parent and exclusive regular-file identities provide that binding).
    A zero-argument predicate cannot authorize an unpredictable descendant and
    therefore fails closed. A temp that changes identity is neither validated,
    published, nor unlinked.

    ``requested_mode`` is mandatory only for a typed comparison result: it must be
    ``formulas``/``values`` for a lone commit or ``both`` when ``twin=True``. This
    explicit signal prevents flavor inference from workbook bytes, filenames, or
    display prose. Untyped legacy producers retain their old behavior."""
    final = Path(final)
    if requested_mode in ("formulas", "values", "both"):
        path_error = _comparison_path_limit_error(final, twin)
        if path_error is not None:
            return ConsolidateResult(status="error", message=path_error)
    validate = validate or (lambda p: _openable_xlsx(p, expect_sheet))
    confirm = confirm_overwrite or (lambda _p: True)
    guard_error = [None]

    def _deny_guard(detail=None):
        if guard_error[0] is None:
            guard_error[0] = (
                "Refusing to finalize the comparison: destination ownership "
                "changed while the comparison was running. No further output "
                "was published; re-run after the destination is stable.")
            if detail:
                log.error("artifact commit boundary rejected: %s", detail)
        return False

    def guard_ok(path, **binding):
        if guard_error[0] is not None:
            return False
        if commit_guard is None:
            return True
        try:
            # New guards consume the exact mutation target and, where supported,
            # the already-created parent identity. Path-only guards remain
            # compatible; a zero-argument predicate cannot authorize the exact
            # unpredictable temp and is intentionally rejected.
            try:
                current = bool(commit_guard(Path(path), **binding))
            except TypeError:  # silent-ok: retry a target-aware guard that lacks binding kwargs
                try:
                    current = bool(commit_guard(Path(path)))
                except TypeError:  # silent-ok: a zero-argument guard is denied, not trusted
                    current = False
        except Exception as e:  # silent-ok: guard defects are logged and fail closed
            first = str(e).splitlines()[0] if str(e) else ""
            log.error("artifact commit guard failed (%s: %s)",
                      type(e).__name__, first)
            current = False
        if not current:
            return _deny_guard(path)
        return current

    final_twin = _values_twin(final) if twin else None
    destinations = [final, final_twin] if twin else [final]
    if (not guard_ok(final.parent)
            or not all(guard_ok(dest) for dest in destinations)):
        return ConsolidateResult(status="error", message=guard_error[0])
    try:
        if captured_sources is None:
            captured_sources = capture_source_identities(source_paths)
        else:
            captured_sources = tuple(captured_sources)
        ensure_outputs_do_not_alias_sources(
            destinations, source_paths, captured_sources=captured_sources,
            require_sources_current=True)
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))
    if (not guard_ok(final.parent)
            or not all(guard_ok(dest) for dest in destinations)):
        return ConsolidateResult(status="error", message=guard_error[0])
    final.parent.mkdir(parents=True, exist_ok=True)
    parent_identity = _plain_entry_identity(final.parent, directory=True)
    if parent_identity is None:
        _deny_guard(f"unprovable output parent: {final.parent}")
        return ConsolidateResult(status="error", message=guard_error[0])

    def parent_current():
        if guard_error[0] is not None:
            return False
        if _plain_entry_identity(final.parent, directory=True) != parent_identity:
            return _deny_guard(f"output parent changed identity: {final.parent}")
        return guard_ok(final.parent, directory_identity=parent_identity)

    def target_current(path):
        return (parent_current()
                and guard_ok(path, anchor_path=final.parent,
                             anchor_identity=parent_identity))

    if not parent_current() or not all(target_current(d) for d in destinations):
        return ConsolidateResult(status="error", message=guard_error[0])
    # Record which destinations existed at the FIRST ask, so the pre-commit TOCTOU
    # re-check (confirm_late_overwrite) only re-prompts for one that APPEARED while
    # produce_fn ran — never for one the user already decided on.
    existed_at_confirm = {}
    for dest in ([final, final_twin] if twin else [final]):
        if not target_current(dest):
            return ConsolidateResult(status="error", message=guard_error[0])
        existed_at_confirm[dest] = dest.exists()
        if existed_at_confirm[dest] and not confirm(dest):
            return ConsolidateResult(status="cancelled",
                                     message="Cancelled. Existing file kept.")

    temp_identities = {}

    def temp_current(path):
        if path is None:
            return True
        path = Path(path)
        expected = temp_identities.get(path)
        if (expected is None
                or _plain_entry_identity(path, directory=False) != expected):
            return _deny_guard(f"producer temp changed identity: {path}")
        return target_current(path)

    def cleanup_temp(path):
        """Unlink only the exact reserved temp under the still-bound parent."""
        if path is None:
            return True
        path = Path(path)
        try:
            path.lstat()
        except (FileNotFoundError, NotADirectoryError):  # silent-ok: proven absence needs no cleanup
            return True
        except OSError:  # silent-ok: identity uncertainty retains the path
            return False
        if not temp_current(path):
            log.error("artifact commit: retained replaced/unowned temp: %s", path)
            return False
        try:
            ensure_outputs_do_not_alias_sources(
                (path,), source_paths, captured_sources=captured_sources)
        except ValueError:
            log.error("artifact commit: retained unsafe temp path instead of "
                      "deleting a selected source: %s", path)
            return False
        try:
            path.unlink()
        except FileNotFoundError:  # silent-ok: concurrent absence is the cleanup goal
            return True
        except OSError:  # silent-ok: failed unlink retains the bound temp
            return False
        try:
            path.lstat()
        except (FileNotFoundError, NotADirectoryError):  # silent-ok: post-unlink absence is success
            return True
        except OSError:  # silent-ok: unverifiable post-state is not claimed clean
            return False
        return False

    def reserve_temp(path):
        """Exclusively create and bind one ordinary producer temp."""
        path = Path(path)
        if not target_current(path):
            return "guard"
        flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
        flags |= getattr(os, "O_BINARY", 0)
        try:
            fd = os.open(path, flags, 0o600)
        except FileExistsError:  # silent-ok: unpredictable collision is untouched; choose a new token
            return "collision"
        except OSError as e:
            log.error("artifact commit: could not reserve %s (%s: %s)",
                      path.name, type(e).__name__, e)
            return "error"
        try:
            identity = _plain_fd_identity(fd)
        finally:
            os.close(fd)
        if (identity is None
                or _plain_entry_identity(path, directory=False) != identity):
            log.error("artifact commit: exclusive temp identity was not stable: %s",
                      path)
            return "unsafe"
        temp_identities[path] = identity
        if not temp_current(path):
            return "guard"
        return "ok"

    tmp = tmp_twin = None
    for _attempt in range(32):
        token = _new_token()
        candidate = _producer_temp(final, token)
        candidate_twin = _values_twin(candidate) if twin else None
        state = reserve_temp(candidate)
        if state == "collision":
            continue
        if state != "ok":
            break
        if candidate_twin is not None:
            twin_state = reserve_temp(candidate_twin)
            if twin_state == "collision":
                if not cleanup_temp(candidate):
                    _deny_guard("could not release a collided twin reservation")
                    break
                temp_identities.pop(candidate, None)
                continue
            if twin_state != "ok":
                cleanup_temp(candidate)
                break
        tmp, tmp_twin = candidate, candidate_twin
        break
    if tmp is None:
        if guard_error[0] is None:
            _deny_guard("could not reserve an exclusive producer temp")
        return ConsolidateResult(status="error", message=guard_error[0])

    # Map BOTH the full temp path AND its basename -> the final equivalents: compare_core's
    # save-error message names only ``path.name`` (the basename), so a full-path-only rewrite
    # would leave the temp NAME in an error message (P2-R02). Full paths first so a basename
    # substring inside a full path is already gone by the time the basename rule runs.
    mapping = {str(tmp): str(final), tmp.name: final.name}
    if twin:
        mapping[str(tmp_twin)] = str(final_twin)
        mapping[tmp_twin.name] = final_twin.name

    # The overwrite prompt and source checks above can be slow. Revalidate at
    # the last callback boundary before the producer opens its temp pathname.
    if not temp_current(tmp) or (twin and not temp_current(tmp_twin)):
        cleanup_temp(tmp)
        cleanup_temp(tmp_twin)
        return ConsolidateResult(status="error", message=guard_error[0])
    try:
        result = produce_fn(tmp)
    except BaseException:
        cleanup_temp(tmp)
        cleanup_temp(tmp_twin)
        raise
    if getattr(result, "status", None) != "ok":
        cleanup_temp(tmp)                 # producer cancelled/errored — nothing to commit
        cleanup_temp(tmp_twin)
        return _rewrite_paths(result, mapping)   # P2-R02: never leak the deleted temp name
    typed_outcome = getattr(result, "comparison_outcome", None)
    if typed_outcome is not None:
        valid_mode = ((twin and requested_mode == "both")
                      or (not twin and requested_mode in ("formulas", "values")))
        valid_outcome = (isinstance(typed_outcome, ComparisonOutcome)
                         and typed_outcome.status == "ok"
                         and typed_outcome.completion in ("complete", "partial"))
        if not valid_outcome or not valid_mode:
            cleanup_temp(tmp)
            cleanup_temp(tmp_twin)
            return ConsolidateResult(
                status="error",
                message=("Could not finalize the comparison: its typed artifact "
                         "mode was missing or inconsistent. No output was published."))
    if not temp_current(tmp) or (twin and not temp_current(tmp_twin)):
        cleanup_temp(tmp)
        cleanup_temp(tmp_twin)
        return ConsolidateResult(status="error", message=guard_error[0])
    # Re-check after production so a destination that appeared as a hardlink to
    # an input during the build is never atomically replaced with the result.
    try:
        ensure_outputs_do_not_alias_sources(
            destinations, source_paths, captured_sources=captured_sources,
            require_sources_current=True)
    except ValueError as e:
        cleanup_temp(tmp)
        cleanup_temp(tmp_twin)
        return ConsolidateResult(status="error", message=str(e))
    # The VALUES workbook is the single transactional artifact (twin), else the lone file.
    primary_tmp, primary_final = (tmp_twin, final_twin) if twin else (tmp, final)
    alias_block = [None]
    schema_block = [None]
    # CMP-AUD-115: a typed comparison's VALUES artifact must also satisfy the
    # versioned comparison-artifact schema before it can replace a good file.
    primary_validate = validate
    if typed_outcome is not None and (twin or requested_mode == "values"):
        def primary_validate(path, _base=validate):
            if not _base(path):
                return False
            problem = comparison_artifact_problem(path, typed_outcome)
            if problem is None:
                return True
            schema_block[0] = (
                f"Could not finalize {primary_final.name}: {problem}. The previous "
                "file (if any) was left unchanged; re-run the comparison.")
            log.error("artifact commit: comparison schema v%d rejected %s — %s",
                      COMPARISON_ARTIFACT_SCHEMA, Path(path).name, problem)
            return False

    def alias_safe(dest):
        def proceed():
            if not target_current(dest):
                alias_block[0] = guard_error[0]
                return False
            try:
                ensure_outputs_do_not_alias_sources(
                    (dest,), source_paths, captured_sources=captured_sources,
                    require_sources_current=True)
                return True
            except ValueError as e:
                alias_block[0] = str(e)
                return False
        return proceed

    # P12 TOCTOU: the single transactional artifact must not silently clobber a file
    # that APPEARED at its destination while produce_fn ran (the user was never asked).
    if not target_current(primary_final) or not temp_current(primary_tmp):
        cleanup_temp(tmp)
        cleanup_temp(tmp_twin)
        return ConsolidateResult(status="error", message=guard_error[0])
    if not confirm_late_overwrite(primary_final, existed_at_confirm[primary_final], confirm):
        cleanup_temp(tmp)
        cleanup_temp(tmp_twin)
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.")
    if not _commit_one(
            primary_tmp, primary_final, primary_validate,
            proceed=alias_safe(primary_final), discard=cleanup_temp,
            temp_current=lambda: temp_current(primary_tmp),
            final_current=lambda: target_current(primary_final)):
        cleanup_temp(tmp)
        cleanup_temp(tmp_twin)
        if alias_block[0] or guard_error[0] or schema_block[0]:
            return ConsolidateResult(
                status="error",
                message=alias_block[0] or guard_error[0] or schema_block[0])
        return ConsolidateResult(
            status="error",
            message=(f"Could not finalize {primary_final.name} — the produced workbook "
                     "was missing/invalid or the destination is open in Excel. The "
                     "previous file (if any) was left unchanged."))

    # A typed comparison generation hashes only the files this transaction
    # actually committed.  In twin mode capture the canonical values member now,
    # before the best-effort formulas branch can lose its guard; normal branches
    # re-hash values immediately before attaching the final generation.
    primary_member = None
    if typed_outcome is not None:
        primary_flavor = "values" if twin else requested_mode
        try:
            primary_member = _committed_artifact_member(
                primary_final, flavor=primary_flavor, commit_role="canonical",
                current_guard=target_current)
        except ValueError as e:
            log.error("comparison generation: %s", e)
            cleanup_temp(tmp)
            cleanup_temp(tmp_twin)
            return ConsolidateResult(
                status="error",
                message=(guard_error[0] or
                         "The committed comparison changed while its identity was "
                         "being recorded; no generation metadata was published."),
                output_path=str(primary_final))
    if twin:
        # The formulas sibling is best-effort. Don't clobber a formulas file that
        # APPEARED during produce either — a decline (like a failed commit) leaves the
        # already-committed values workbook as the canonical, truthful output.
        if not target_current(final) or not temp_current(tmp):
            cleanup_temp(tmp)
            alias_block[0] = guard_error[0]
            committed_formulas = False
        elif not confirm_late_overwrite(final, existed_at_confirm[final], confirm):
            cleanup_temp(tmp)
            committed_formulas = False
        else:
            alias_block[0] = None
            committed_formulas = _commit_one(
                tmp, final, validate, proceed=alias_safe(final),
                discard=cleanup_temp,
                temp_current=lambda: temp_current(tmp),
                final_current=lambda: target_current(final))
        if not committed_formulas:
            if guard_error[0] is not None:
                rewritten = _rewrite_paths(result, mapping)
                terminal = ConsolidateResult(
                    status="error", message=guard_error[0],
                    output_path=str(final_twin),
                    summary_lines=list(getattr(rewritten, "summary_lines", ()) or ()),
                    verdict=getattr(rewritten, "verdict", None),
                    completion=getattr(rewritten, "completion", None),
                    skipped_inputs=getattr(rewritten, "skipped_inputs", 0),
                    failed_inputs=getattr(rewritten, "failed_inputs", 0),
                    comparison_outcome=typed_outcome)
                if primary_member is not None:
                    _attach_artifact_generation(
                        terminal, (primary_member,), requested_mode)
                    terminal.attempt_state = AttemptState(
                        state="failed", message=guard_error[0],
                        generation_id=terminal.artifact_generation.generation_id)
                return _publish_artifact_generation(terminal, commit_guard)
            log.warning("comparison: the live-formulas workbook for %s was not finalized; the "
                        "values workbook is committed", final.name)
            result = _rewrite_paths(result, mapping)
            # P2-R03: be TRUTHFUL — the formulas file was not written. Point output_path at the
            # committed values workbook and turn the formulas line into a not-refreshed warning
            # (any pre-existing formulas file at `final` is stale, not this comparison).
            result.output_path = str(final_twin)
            result.summary_lines = [
                ("Live-formulas file: NOT refreshed (best-effort write failed; the values "
                 "workbook above is the canonical output)")
                if s.startswith("Live-formulas file:") else s
                for s in (result.summary_lines or [])]
            if typed_outcome is not None:
                try:
                    primary_member = _committed_artifact_member(
                        final_twin, flavor="values", commit_role="canonical",
                        current_guard=target_current)
                except ValueError as e:
                    log.error("comparison generation: %s", e)
                    return ConsolidateResult(
                        status="error",
                        message=(guard_error[0] or
                                 "The committed values comparison changed while its "
                                 "identity was being recorded; no generation metadata "
                                 "was published."),
                        output_path=str(final_twin))
                _attach_artifact_generation(
                    result, (primary_member,), requested_mode)
            return _publish_artifact_generation(result, commit_guard)
        if typed_outcome is not None:
            try:
                # Refresh the values binding at the final generation boundary,
                # then bind the best-effort formulas member that did commit.
                primary_member = _committed_artifact_member(
                    final_twin, flavor="values", commit_role="canonical",
                    current_guard=target_current)
                formulas_member = _committed_artifact_member(
                    final, flavor="formulas", commit_role="best_effort",
                    current_guard=target_current)
            except ValueError as e:
                log.error("comparison generation: %s", e)
                return ConsolidateResult(
                    status="error",
                    message=(guard_error[0] or
                             "A committed comparison member changed while generation "
                             "metadata was being recorded; no generation metadata was "
                             "published."),
                    output_path=str(final_twin))
            result = _rewrite_paths(result, mapping)
            _attach_artifact_generation(
                result, (primary_member, formulas_member), requested_mode)
            return _publish_artifact_generation(result, commit_guard)
    result = _rewrite_paths(result, mapping)
    if typed_outcome is not None:
        _attach_artifact_generation(result, (primary_member,), requested_mode)
    return _publish_artifact_generation(result, commit_guard)


# --------------------------------------------------------------------------- #
# input fingerprint (F5 / R1-R03)
# --------------------------------------------------------------------------- #
def _is_excluded(name):
    if name.startswith("~$"):                       # Excel lock file
        return True
    if ".tmp-" in name or name.endswith(".tmp"):    # our in-flight temp
        return True
    if _COMPARISON_PAYLOAD_RE.fullmatch(name):       # strict v3 comparison metadata
        return True
    if name == _COMPARISON_PUBLICATION_LOCK_NAME:     # permanent transaction lease anchor
        return True
    return any(name.endswith(s) for s in _FP_EXCLUDED_SUFFIXES)


def is_report_data_file(name):
    """True iff `name` is a real per-route / consolidated REPORT data file — a
    ``.xlsx`` or ``.pdf`` (`_REPORT_SUFFIXES`) that is NOT an Excel lock (``~$``),
    an in-flight temp, a comparison payload, the publication lock, or one of our
    sidecars (`_is_excluded`).

    The ONE accepted-data-file predicate that Matrix presence, newest-data mtime,
    fingerprinting, and adapter discovery share (CMP-AUD-083). A folder holding
    only ``~$route.xlsx``, ``notes.txt``, ``README``, ``.fingerprint.json``, or a
    newer lock therefore no longer reads as an export / a fresher artifact. Purely
    name-based; callers still confirm ``is_file()`` for real directory entries."""
    return name.lower().endswith(_REPORT_SUFFIXES) and not _is_excluded(name)


# --------------------------------------------------------------------------- #
# CMP-AUD-080 — CONTENT identity for every effective source.
#
# The v1 fingerprint hashed (name, size, mtime_ns), so replacing a file with
# different same-length bytes and restoring its timestamp left the cached
# "match / 0 differences" fresh. v2 hashes the BYTES.
#
# Re-reading a statewide store on every snapshot would be unaffordable, so each
# file's digest is memoized against a CHANGE TOKEN — never against stat alone,
# which is exactly the memoization the audit prohibits. On Windows the token
# includes the filesystem's own ChangeTime (FILE_BASIC_INFO.ChangeTime), which
# the OS advances on any write to the file's data or metadata and which
# SetFileTime cannot restore: measured here, an in-place same-size rewrite with
# the mtime put back leaves (size, mtime_ns, file id) byte-identical while
# ChangeTime moves. A whole-file replacement moves the file id as well. Where no
# change token can be obtained (a non-Windows filesystem, an unreadable handle),
# the memo is refused and the file is re-hashed — fail-safe, never fail-fast.
# --------------------------------------------------------------------------- #
_HASH_BLOCK = 1 << 20
_DIGEST_MEMO = {}
_DIGEST_MEMO_MAX = 50_000          # ~6 statewide environments' worth of routes
# FILETIME (100 ns ticks since 1601-01-01) -> unix nanoseconds.
_FILETIME_EPOCH_TICKS = 116_444_736_000_000_000
# A change token can only be as fine as the system clock that stamps it: two
# writes inside one tick get the SAME ChangeTime (measured ~10 ms here, coarser
# on some virtualized hosts — a CI runner produced an unmoved token). So a digest
# is memoized only once its file's change stamp is comfortably in the PAST: a
# file changed within this window is re-hashed next time instead of trusted.
# (The same "racily clean" rule Git applies to its stat cache.) One second is far
# above any plausible tick, and the cost is only re-hashing a just-written file.
_RACY_WINDOW_NS = 1_000_000_000


def _change_token(path, st):
    """A tamper-visible identity for one file, or None when none is available."""
    base = (int(st.st_size), int(getattr(st, "st_mtime_ns", 0)),
            int(getattr(st, "st_dev", 0)), int(getattr(st, "st_ino", 0)))
    if os.name != "nt":
        return None                # no verified change signal -> never memoize
    change_time = _windows_change_time(path)
    return None if change_time is None else base + (change_time,)


_WIN32_METADATA = None


def _win32_metadata_api():
    """``(ctypes, kernel32, FILE_BASIC_INFO, INVALID_HANDLE)`` or None. Built once,
    lazily: ctypes stays off every non-Windows and non-fingerprint path."""
    global _WIN32_METADATA
    if _WIN32_METADATA is None:
        _WIN32_METADATA = False
        if os.name == "nt":
            try:
                import ctypes
                from ctypes import wintypes
                k32 = ctypes.WinDLL("kernel32", use_last_error=True)
                k32.CreateFileW.restype = wintypes.HANDLE
                k32.CreateFileW.argtypes = [
                    wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD,
                    ctypes.c_void_p, wintypes.DWORD, wintypes.DWORD,
                    wintypes.HANDLE]
                k32.GetFileInformationByHandleEx.restype = wintypes.BOOL
                k32.GetFileInformationByHandleEx.argtypes = [
                    wintypes.HANDLE, ctypes.c_int, ctypes.c_void_p, wintypes.DWORD]
                k32.CloseHandle.argtypes = [wintypes.HANDLE]

                class FILE_BASIC_INFO(ctypes.Structure):
                    _fields_ = [("CreationTime", ctypes.c_longlong),
                                ("LastAccessTime", ctypes.c_longlong),
                                ("LastWriteTime", ctypes.c_longlong),
                                ("ChangeTime", ctypes.c_longlong),
                                ("FileAttributes", wintypes.DWORD)]

                invalid = ctypes.cast(wintypes.HANDLE(-1), ctypes.c_void_p).value
                _WIN32_METADATA = (ctypes, k32, FILE_BASIC_INFO, invalid)
            except (OSError, AttributeError, ValueError) as e:
                log.warning("content identity: no Windows change token available "
                            "(%s: %s); source digests will be recomputed",
                            type(e).__name__, e)
    return _WIN32_METADATA or None


def _windows_change_time(path):
    """FILE_BASIC_INFO.ChangeTime for `path` (None when it can't be read).

    Measured cheaper than ``os.stat`` (one metadata handle, no path
    re-resolution), so validating the memo costs nothing next to the v1 stat walk
    it replaces."""
    api = _win32_metadata_api()
    if api is None:
        return None
    ctypes, k32, basic_info, invalid = api
    # 0 access + share-all + FILE_FLAG_BACKUP_SEMANTICS: metadata only, never
    # blocking a workbook another process has open.
    handle = k32.CreateFileW(str(path), 0, 7, None, 3, 0x02000000, None)
    if not handle or handle == invalid:
        return None
    try:
        info = basic_info()
        if not k32.GetFileInformationByHandleEx(
                handle, 0, ctypes.byref(info), ctypes.sizeof(info)):
            return None
        return int(info.ChangeTime)
    finally:
        k32.CloseHandle(handle)


def _memoizable(token):
    """True when `token`'s change stamp is old enough to be trusted (see
    ``_RACY_WINDOW_NS``). A token with no change stamp is never memoizable."""
    if token is None:
        return False
    changed_ns = (token[-1] - _FILETIME_EPOCH_TICKS) * 100
    return (time.time_ns() - changed_ns) > _RACY_WINDOW_NS


def content_digest(path):
    """The SHA-256 of `path`'s bytes, memoized against its change token.

    The memo is written only when the token is BOTH unchanged across the read
    (so a file edited mid-hash is never cached) and old enough to be outside the
    clock's own granularity (so a same-tick rewrite can never hide behind it).
    Everything else re-hashes.

    The token is always taken from ``os.stat``: a ``DirEntry.stat()`` from a
    directory walk reports ``st_dev``/``st_ino`` as 0 on Windows, so a token
    built from one could never compare equal to a token built from the other —
    the file identity has to come from the same source on both sides.

    Raises OSError when the file cannot be read — callers turn that into the
    ``_UNREADABLE`` fingerprint rather than a silent match."""
    path = Path(path)
    token = _change_token(path, path.stat())
    key = os.path.normcase(str(path))
    if token is not None:
        cached = _DIGEST_MEMO.get(key)
        if cached is not None and cached[0] == token:
            return cached[1]
    digest = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            block = f.read(_HASH_BLOCK)
            if not block:
                break
            digest.update(block)
    value = digest.hexdigest()
    if token is not None and _memoizable(token):
        try:
            after = _change_token(path, path.stat())
        except OSError:  # silent-ok: an unstattable file is simply not memoized
            after = None
        if after == token:
            if len(_DIGEST_MEMO) >= _DIGEST_MEMO_MAX:
                _DIGEST_MEMO.clear()   # bounded: a cleared memo only costs a re-hash
            _DIGEST_MEMO[key] = (token, value)
    return value


def fingerprint(folder):
    """A stable CONTENT identity string over the DATA files directly inside
    `folder`: a hash of the sorted ``(name, sha256-of-bytes)`` pairs plus the file
    count. Catches a file added / removed / resized / re-timed AND a same-size,
    same-timestamp replacement of its bytes (CMP-AUD-080) — the case the v1
    ``(name, size, mtime_ns)`` fingerprint could not see. Excludes Excel lock
    files, our own sidecars (``.fingerprint.json`` / ``.outcome.json``), and
    in-flight temp / ``.staging`` siblings; sub-directories are ignored (stores
    are flat per-route folders). An unreadable folder or file yields the
    ``_UNREADABLE`` sentinel so the caller rebuilds rather than silently
    matching. Never raises.

    The v1 -> v2 schema change makes every fingerprint recorded by an older build
    compare unequal, so each cell and consolidated workbook reads stale exactly
    ONCE and rebuilds against content identity (the required metadata-only
    migration)."""
    folder = Path(folder)
    try:
        # scandir (not iterdir): the name filter and is_file() test come from
        # the directory read already performed. The digest's own change token is
        # taken from os.stat inside content_digest — a DirEntry stat has no file
        # identity on Windows, so it cannot serve as one side of that comparison.
        with os.scandir(folder) as it:
            entries = sorted(it, key=lambda e: e.name)
    except OSError:
        return _UNREADABLE
    parts = []
    for e in entries:
        # A CHANGE-detection hash is deliberately conservative-INCLUSION: it counts
        # every file except the strict-format artifacts WE create (locks, temps,
        # exact-format comparison payloads, sidecars). A near-match `.zlib` or a
        # stray user file MUST still participate so nothing hides from freshness —
        # this is a stronger property than `is_report_data_file`'s "is this an
        # export?" allowlist (which powers presence/mtime/discovery, CMP-AUD-083),
        # so fingerprint keeps `_is_excluded`, NOT the extension allowlist.
        if _is_excluded(e.name):
            continue
        try:
            if not e.is_file():
                continue
            parts.append(f"{e.name}\0{content_digest(e.path)}")
        except OSError:
            return _UNREADABLE
    digest = hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()[:32]
    return f"v{_FP_SCHEMA}:{len(parts)}:{digest}"


def _fp_sidecar(consolidated):
    return Path(str(consolidated) + _FP_SUFFIX)


# A fingerprint value that can NEVER equal a real one (`fingerprint()` returns
# ``v<schema>:<count>:<hexdigest>`` or ``_UNREADABLE``), so a sidecar carrying it always
# reads STALE — used to fail-safe an undeletable stale sidecar (P2-A02).
_FP_RACE_SENTINEL = "__race_invalidated__"


def _write_fp_sentinel(p):
    """Overwrite sidecar `p` with the non-matching race sentinel, so the workbook reads stale
    even when the old sidecar could not be removed (P2-A02). Best-effort; True iff written."""
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump({"schema_version": _FP_SCHEMA, "fingerprint": _FP_RACE_SENTINEL}, f)
        return True
    except OSError:
        return False


def _quarantine_workbook(consolidated, commit_guard=None):
    """LAST-resort A02 fail-safe: when a stale sidecar can be NEITHER removed nor overwritten,
    rename the freshly-replaced (race-suspect) workbook aside so the canonical path resolves
    MISSING — `consolidated_fresh` then reads stale (the workbook is absent) and a clean
    rebuild happens. Best-effort; True iff the workbook is no longer at its canonical path."""
    consolidated = Path(consolidated)
    q = consolidated.with_name(consolidated.name + ".race-stale")
    try:
        if commit_guard is not None and not commit_guard(q):
            return False
        _silent_unlink(q)                            # clear a prior quarantine so rename won't clash
        if commit_guard is not None and (not commit_guard(consolidated)
                                         or not commit_guard(q)):
            return False
        consolidated.rename(q)
        return True
    except OSError:
        return not consolidated.exists()             # treat an already-gone workbook as quarantined


def write_consolidated_fingerprint(consolidated, store_dir, built_from=None,
                                   commit_guard=None):
    """Record `store_dir`'s input fingerprint beside `consolidated` after a successful
    build, so a later REUSE can tell whether the inputs changed (R1-R03). Atomic +
    best-effort: a write failure just means the next reuse rebuilds (fail-safe — a
    missing/unwritable sidecar reads as stale). Returns True iff the sidecar was written.

    `built_from` (P2-A02): the fingerprint captured BEFORE the build. If it differs from
    the current (post-build) fingerprint the inputs changed DURING the build, so the
    workbook may reflect a pre-change or mixed input set — do NOT publish a 'fresh'
    sidecar (the workbook reads stale and rebuilds next time). The GUI task lock already
    serializes writers; this guards an external mid-build mutation.

    `commit_guard(path)` is an optional target-aware ownership predicate. It is
    called with each exact temp/final path immediately before mutation. False or
    an exception fails closed; a temp is cleaned only while its path is allowed.
    """
    consolidated = Path(consolidated)

    def _allowed(path, action):
        if commit_guard is None:
            return True
        try:
            current = bool(commit_guard(Path(path)))
        except Exception as e:  # noqa: BLE001 â€” sidecar publication is best-effort
            log.warning("artifact fingerprint for %s: ownership check failed before %s "
                        "(%s: %s); retained", consolidated.name, action,
                        type(e).__name__, e)
            return False
        if not current:
            log.warning("artifact fingerprint for %s: destination changed before %s; retained",
                        consolidated.name, action)
        return current

    fp = fingerprint(store_dir)
    if built_from is not None and built_from != fp:
        # P2-A02: the producer ALREADY replaced the workbook, so a leftover sidecar from a PRIOR
        # build could still certify it fresh (esp. if the inputs reverted to the old identity).
        # REMOVE the stale sidecar; if it can't be removed (locked), OVERWRITE it with a
        # guaranteed-non-matching sentinel so consolidated_fresh still reads stale. Log the
        # ACTUAL outcome (P2-A02) — never leave a freshly-replaced workbook reading fresh.
        # Fail-safe ladder (each rung only reached if the prior failed): remove the stale
        # sidecar -> overwrite it with a non-matching sentinel -> (P2-A02 dual-failure)
        # quarantine the replaced workbook so the canonical path resolves MISSING -> log
        # critically. The replaced workbook must NEVER stay eligible to read fresh.
        sc = _fp_sidecar(consolidated)
        if not _allowed(sc, "stale-sidecar invalidation"):
            log.error("artifact fingerprint for %s could not invalidate stale state because "
                      "the destination ownership changed", consolidated.name)
        elif _silent_unlink(sc):
            log.warning("artifact fingerprint for %s NOT recorded — inputs changed during the "
                        "build; stale sidecar removed so the workbook rebuilds next reuse",
                        consolidated.name)
        elif (_allowed(sc, "stale-sidecar sentinel write")
              and _write_fp_sentinel(sc)):
            log.warning("artifact fingerprint for %s NOT recorded — inputs changed mid-build and "
                        "the stale sidecar could not be removed; wrote a non-matching sentinel so "
                        "the workbook rebuilds next reuse", consolidated.name)
        elif (_allowed(consolidated, "workbook quarantine")
              and _quarantine_workbook(
                  consolidated,
                  commit_guard=lambda path: _allowed(path, "workbook quarantine"))):
            log.warning("artifact fingerprint for %s NOT recorded — the stale sidecar could be "
                        "neither removed nor overwritten; quarantined the replaced workbook so it "
                        "rebuilds next reuse", consolidated.name)
        else:
            log.critical("artifact fingerprint for %s NOT recorded and neither its stale sidecar "
                         "nor the replaced workbook could be invalidated — it may read fresh; "
                         "re-run", consolidated.name)
        return False
    p = _fp_sidecar(consolidated)
    tmp = p.with_name(p.name + f".tmp-{_new_token()}")
    if not _allowed(tmp, "sidecar temp creation"):
        return False
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"schema_version": _FP_SCHEMA, "fingerprint": fp}, f)
        if (not _allowed(tmp, "sidecar temp commit")
                or not _allowed(p, "sidecar final commit")):
            raise OSError("fingerprint destination ownership changed before commit")
        os.replace(tmp, p)
        return _allowed(p, "sidecar post-commit validation")
    except OSError as e:
        log.warning("artifact fingerprint for %s not written (%s: %s); next reuse rebuilds",
                    consolidated.name, type(e).__name__, e)
        if _allowed(tmp, "sidecar temp cleanup"):
            _silent_unlink(tmp)
        return False


def _read_fp_sidecar(consolidated):
    """The recorded fingerprint string, or None when the sidecar is absent / unreadable
    / corrupt / a different schema. Never raises."""
    try:
        with open(_fp_sidecar(consolidated), encoding="utf-8") as f:
            meta = json.load(f)
    except (OSError, ValueError):
        return None
    if (not isinstance(meta, dict) or meta.get("schema_version") != _FP_SCHEMA
            or not isinstance(meta.get("fingerprint"), str)):
        return None
    return meta["fingerprint"]


def consolidated_fresh(consolidated, store_dir):
    """True iff `consolidated` exists AND its recorded input fingerprint matches
    `store_dir`'s CURRENT fingerprint. A missing workbook, a missing/corrupt/old
    sidecar, an unreadable store, or a fingerprint MISMATCH ⇒ False (rebuild). This is
    identity-based freshness (F5): a DELETED route changes the fingerprint even though
    the newest-mtime is untouched. A legacy workbook with no sidecar reads stale ONCE,
    rebuilds, and records the sidecar (the one-time migration). Never raises."""
    if not Path(consolidated).exists():
        return False
    recorded = _read_fp_sidecar(consolidated)
    if recorded is None:
        return False
    current = fingerprint(store_dir)
    if current == _UNREADABLE:
        return False
    return recorded == current


# --------------------------------------------------------------------------- #
# journaled store promotion + startup recovery (F2)
# --------------------------------------------------------------------------- #
_PROMOTE_DIRNAME = ".promote"
_BAK_INFIX = ".bak-"
_STAGING_SUFFIX = ".staging"


def _plain_entry_identity(path, *, directory):
    """Replacement-sensitive identity for a non-reparse directory/file."""
    try:
        st = Path(path).lstat()
    except OSError:  # silent-ok: an unprovable journal boundary is rejected
        return None
    attrs = getattr(st, "st_file_attributes", 0)
    reparse = getattr(statmod, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    if statmod.S_ISLNK(st.st_mode) or bool(attrs & reparse):
        return None
    expected = statmod.S_ISDIR if directory else statmod.S_ISREG
    if not expected(st.st_mode) or not st.st_ino:
        return None
    return (st.st_dev, st.st_ino)


def _plain_fd_identity(fd):
    """Identity of an already-open ordinary file descriptor, or ``None``.

    A pathname can be replaced between ``lstat`` and ``open``.  Journal readers
    compare this descriptor identity with the previously captured *non-reparse*
    pathname identity before consuming a byte, so opening a swapped symlink or a
    file reached through a replaced journal directory fails closed.
    """
    try:
        st = os.fstat(fd)
    except OSError:  # silent-ok: an unprovable descriptor is rejected
        return None
    if not statmod.S_ISREG(st.st_mode) or not st.st_ino:
        return None
    return (st.st_dev, st.st_ino)


def _entry_present_lstat(path):
    """Whether a directory entry exists without following a reparse point.

    ``Path.exists()`` is false for a broken symlink/junction.  Recovery must not
    mistake that still-present residue for absence and discard its journal.
    Unusual lstat failures are treated conservatively as "present".
    """
    try:
        Path(path).lstat()
        return True
    except (FileNotFoundError, NotADirectoryError):  # silent-ok: proven absent
        return False
    except OSError:  # silent-ok: uncertainty is not proof that cleanup completed
        return True


def _entry_absent_lstat(path):
    return not _entry_present_lstat(path)


def _plain_entry_current(path, identity, *, directory):
    return identity is not None and _plain_entry_identity(
        path, directory=directory) == identity


def _plain_journal_dir_identity(jdir):
    return _plain_entry_identity(jdir, directory=True)


def _journal_dir_current(jdir, identity):
    return identity is not None and _plain_journal_dir_identity(jdir) == identity


def _plain_journal_identity(jdir, jdir_identity, journal):
    """Identity of one direct, regular, non-reparse JSON entry in a bound dir."""
    jdir, journal = Path(jdir), Path(journal)
    if journal.parent != jdir or not _journal_dir_current(jdir, jdir_identity):
        return None
    identity = _plain_entry_identity(journal, directory=False)
    if identity is None or not _journal_dir_current(jdir, jdir_identity):
        return None
    return identity


def _journal_current(jdir, jdir_identity, journal, journal_identity):
    return (_plain_journal_identity(jdir, jdir_identity, journal)
            == journal_identity)


_JOURNAL_INVALID = object()
_JOURNAL_UNSAFE = object()
_ABSENT_ENTRY = object()


def _read_bound_journal(jdir, jdir_identity, journal, journal_identity):
    """Read one JSON journal through an identity-verified file descriptor.

    ``_JOURNAL_INVALID`` means the same bound ordinary entry was unreadable or
    malformed (recovery may drop that junk in an owned store).
    ``_JOURNAL_UNSAFE`` means the directory/file identity changed; callers must
    leave every pathname untouched.  Crucially, the descriptor identity is
    checked *before* ``fdopen``/JSON parsing, closing the lstat-to-open read escape.
    """
    if not _journal_current(jdir, jdir_identity, journal, journal_identity):
        return _JOURNAL_UNSAFE
    flags = os.O_RDONLY | getattr(os, "O_BINARY", 0)
    flags |= getattr(os, "O_NOINHERIT", 0) | getattr(os, "O_NOFOLLOW", 0)
    fd = None
    try:
        fd = os.open(journal, flags)
        if (_plain_fd_identity(fd) != journal_identity
                or not _journal_current(jdir, jdir_identity,
                                        journal, journal_identity)):
            return _JOURNAL_UNSAFE
        with os.fdopen(fd, "r", encoding="utf-8") as f:
            fd = None
            try:
                value = json.load(f)
            except (OSError, ValueError):  # silent-ok: caller distinguishes owned junk from an unsafe replacement
                value = _JOURNAL_INVALID
        if not _journal_current(jdir, jdir_identity, journal, journal_identity):
            return _JOURNAL_UNSAFE
        return value
    except OSError:  # silent-ok: open failure is classified by the still-bound pathname
        return (_JOURNAL_INVALID
                if _journal_current(jdir, jdir_identity, journal, journal_identity)
                else _JOURNAL_UNSAFE)
    finally:
        if fd is not None:
            try:
                os.close(fd)
            except OSError:  # silent-ok: best-effort close of a rejected descriptor
                pass


def _unlink_plain_file(path, expected_identity):
    """Best-effort unlink of the exact captured regular non-reparse entry."""
    path = Path(path)
    if not _plain_entry_current(path, expected_identity, directory=False):
        return False
    try:
        path.unlink()
    except FileNotFoundError:  # silent-ok: disappearance means identity changed; retain
        return False                         # it changed after the identity check
    except OSError:  # silent-ok: locked/unlinkable exact journal is retained for retry
        return False
    return _entry_absent_lstat(path)


def _unlink_bound_journal(jdir, jdir_identity, journal, journal_identity):
    """Unlink only the captured journal inside the captured ordinary directory."""
    if not _journal_current(jdir, jdir_identity, journal, journal_identity):
        return False
    removed = _unlink_plain_file(journal, journal_identity)
    return removed and _journal_dir_current(jdir, jdir_identity)


def _ensure_plain_journal_dir(jdir):
    """Create `.promote` itself, or bind an existing ordinary directory.

    `exist_ok=True` follows a planted junction.  Create exclusively instead,
    then lstat-bind the exact ordinary directory identity in either branch.
    """
    jdir = Path(jdir)
    try:
        jdir.mkdir(exist_ok=False)
    except FileExistsError:  # silent-ok: an existing entry is lstat-validated below
        pass
    except OSError:  # silent-ok: inability to create the boundary fails promotion closed
        return None
    return _plain_journal_dir_identity(jdir)


def _journal_dir(live):
    """The promotion journal dir, in the destination PARENT — never inside `live`
    (which is itself renamed), per §C.2."""
    return Path(live).parent / _PROMOTE_DIRNAME


def _rmtree(path, expected_identity=None):
    """Remove only an ordinary, non-reparse directory of the expected identity.

    ``shutil.rmtree`` itself avoids following symlink children on supported
    platforms, but its root pathname still needs an lstat identity binding.  A
    junction, broken reparse point, or replacement directory is retained.
    """
    import shutil
    path = Path(path)
    identity = _plain_entry_identity(path, directory=True)
    if identity is None:
        return _entry_absent_lstat(path)
    if expected_identity is not None and identity != expected_identity:
        return False
    if not _plain_entry_current(path, identity, directory=True):
        return False
    shutil.rmtree(path, ignore_errors=True)
    return _entry_absent_lstat(path)


def _rmtree_gone(path, expected_identity=None):
    """rmtree `path` (best-effort) and report whether it is GONE afterwards — True when the
    path no longer exists (removed, or never there), False when it survives (locked). Lets
    recovery observe cleanup success so it can retain the journal on a failed cleanup (P2-R05)."""
    path = Path(path)
    if _entry_absent_lstat(path):
        return True
    identity = _plain_entry_identity(path, directory=True)
    if identity is None:
        return False                          # includes broken symlinks/junctions
    if expected_identity is not None and identity != expected_identity:
        return False
    _rmtree(path, identity)
    return _entry_absent_lstat(path)


def _finalize_journal(journal, *residue, guard=None,
                      residue_identities=None, unlink_journal=None):
    """Remove `journal` ONLY after EVERY named residue dir is confirmed GONE (P2-R05) — so a
    locked backup/staging RETAINS the journal and the next launch's recovery retries the
    cleanup. Used by BOTH promote_store's completion/restore branches and recovery. Startup
    recovery supplies ``guard`` so the captured store identity is revalidated immediately
    before every residue/journal mutation. Returns True iff the journal was removed."""
    all_gone = True
    for path in residue:
        if guard is not None and not guard():
            return False
        expected = None
        if residue_identities is not None:
            expected = residue_identities.get(Path(path))
            if expected is _ABSENT_ENTRY:
                if _entry_absent_lstat(path):
                    continue
                all_gone = False                    # a late same-name entry is not ours
                break
        if not _rmtree_gone(path, expected):
            all_gone = False
            break
    if all_gone:
        if guard is not None and not guard():
            return False
        return (unlink_journal() if unlink_journal is not None
                else _silent_unlink(journal))
    log.warning("promotion: residue cleanup incomplete; journal %s RETAINED for next-launch "
                "retry", Path(journal).name)
    return False


def _is_report_artifact(name):
    """A real per-route report file (P2-R01): a non-excluded regular name ending in a
    report suffix (.xlsx / .pdf). A `.txt` or other foreign file does NOT qualify."""
    return (not _is_excluded(name)) and name.lower().endswith(_REPORT_SUFFIXES)


def _staging_has_report_file(staged):
    """A staging dir is committable only if it exists and holds at least one REPORT artifact
    (.xlsx/.pdf) — not merely any regular file (P2-R01: a staging containing only a nested
    directory, only lock/temp/sidecar files, or only a foreign `.txt` must never replace a
    good live). Excludes Excel locks, our temp/`.staging`/sidecar names, and sub-directories."""
    try:
        staged = Path(staged)
        if _plain_entry_identity(staged, directory=True) is None:
            return False
        for e in staged.iterdir():
            try:
                if (_plain_entry_identity(e, directory=False) is not None
                        and _is_report_artifact(e.name)):
                    return True
            except OSError:
                continue
        return False
    except OSError:
        return False


def is_usable_store(d):
    """True iff `d` is a directory that genuinely holds a report artifact — a real live
    store, not an empty placeholder or a foreign directory. Recovery treats ONLY this as
    proof that the canonical copy exists (P2-B02: a bare `exists()` is not proof); the
    worker uses it too, so a failed-promotion artifact is truthful (P2-A04)."""
    return (_plain_entry_identity(d, directory=True) is not None
            and _staging_has_report_file(d))


def _dir_is_disposable_placeholder(d):
    """True iff `d` is a directory whose every entry is disposable (an excluded temp / lock /
    sidecar name) — safe to remove when restoring a real backup over it. A directory with
    ANY sub-directory or non-excluded (incl. foreign) file is NOT disposable: that is a
    genuine conflict the recovery must NOT silently destroy (P2-B02)."""
    try:
        d = Path(d)
        if not d.is_dir():
            return False
        for e in d.iterdir():
            # The fixed publication lock is intentionally permanent and may be
            # held by another process. Never make its parent recursively
            # disposable merely because fingerprinting excludes the lock.
            if e.name == _COMPARISON_PUBLICATION_LOCK_NAME:
                return False
            if e.is_dir() or not _is_excluded(e.name):
                return False
        return True
    except OSError:
        return False


def promote_store(live, staged, guard=None):
    """Replace `live` with the freshly-built `staged` folder, JOURNALED so a crash between
    the renames is recoverable (F2: never zero copies, never a truncated last-good). Steps
    (§C.2): validate staging holds a report artifact -> write the journal in the destination
    parent -> rename(live -> live.bak-<token>) -> rename(staged -> live) -> delete the
    journal -> drop the backup. A failure mid-swap restores `live` from the backup and
    discards `staged`; a blocked restore RETAINS the journal + backup for next-launch
    recovery. The FIRST promotion (no prior `live`) is journaled too, so a failed rename
    retains the staging + journal rather than deleting the only completed copy (P2-B06).
    ``guard`` is an optional no-argument ownership-lease predicate; when supplied it is
    rechecked immediately before every mutation/cleanup. Returns True iff `live` now holds
    the promoted content under that still-current lease. Never raises."""
    live, staged = Path(live), Path(staged)
    parent_identity = _plain_entry_identity(live.parent, directory=True)
    if (parent_identity is None
            or staged.parent != live.parent):
        log.error("store promotion for %s: destination parent is unsafe or staging is "
                  "outside it; retained", live.name)
        return False

    def _allowed(action):
        if guard is None:
            return True
        try:
            current = bool(guard())
        except Exception as e:  # noqa: BLE001 — promotion remains best-effort
            log.warning("store promotion for %s: ownership check failed before %s "
                        "(%s: %s); retained", live.name, action,
                        type(e).__name__, e)
            return False
        if not current:
            log.warning("store promotion for %s: owned root changed before %s; "
                        "journal and copies retained", live.name, action)
        return current

    staged_identity = _plain_entry_identity(staged, directory=True)
    if (staged_identity is None or not _staging_has_report_file(staged)
            or not _plain_entry_current(staged, staged_identity, directory=True)):
        log.warning("store promotion for %s: staging missing/empty/no-report-file — kept "
                    "last-good", live.name)
        # Empty ordinary staging is our disposable residue. A symlink/junction,
        # broken reparse point, or replacement is never followed or removed.
        if staged_identity is not None and _allowed("invalid-staging cleanup"):
            _rmtree(staged, staged_identity)
        return False

    had_prior = _entry_present_lstat(live)
    live_identity = (_plain_entry_identity(live, directory=True)
                     if had_prior else None)
    if had_prior and live_identity is None:
        log.error("store promotion for %s: live path is not an ordinary directory; "
                  "live and staging retained", live.name)
        return False
    token = _new_token()
    backup = live.with_name(f"{live.name}{_BAK_INFIX}{token}")
    if _entry_present_lstat(backup):
        log.error("store promotion for %s: backup collision at %s; live and staging retained",
                  live.name, backup.name)
        return False
    jdir = _journal_dir(live)
    journal = jdir / f"{token}.json"
    if (not _allowed("journal directory creation")
            or not _plain_entry_current(live.parent, parent_identity, directory=True)):
        return False
    jdir_identity = _ensure_plain_journal_dir(jdir)
    if (jdir_identity is None
            or not _plain_entry_current(live.parent, parent_identity, directory=True)):
        log.error("store promotion for %s: refusing unsafe/reparse journal directory %s; "
                  "live and staging retained", live.name, jdir)
        return False
    # P2-B07: a DURABLE, monotonic transaction generation — strictly greater than every existing
    # same-target journal (derived from the on-disk journals, NOT wall-clock), so recovery's
    # newest-`gen`-first ordering cannot be inverted by a system-clock regression across restarts.
    gen = _next_generation(jdir, live.name, jdir_identity)
    if gen is None:
        log.error("store promotion for %s: journal directory changed during generation scan; "
                  "live and staging retained", live.name)
        return False
    created_journal = False
    journal_identity = None
    fd = None
    try:
        if not _allowed("journal write"):
            return False
        if (not _plain_entry_current(live.parent, parent_identity, directory=True)
                or not _journal_dir_current(jdir, jdir_identity)):
            log.error("store promotion for %s: journal directory changed before write; retained",
                      live.name)
            return False
        flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
        flags |= getattr(os, "O_BINARY", 0) | getattr(os, "O_NOINHERIT", 0)
        fd = os.open(journal, flags, 0o600)
        created_journal = True
        fd_identity = _plain_fd_identity(fd)
        journal_identity = _plain_journal_identity(
            jdir, jdir_identity, journal)
        if (fd_identity is None or journal_identity != fd_identity
                or not _journal_dir_current(jdir, jdir_identity)):
            raise OSError("journal boundary changed during exclusive create")
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            fd = None
            json.dump({"target": live.name, "backup": backup.name,
                       "staging": staged.name, "token": token, "gen": gen}, f)
            f.flush()
            os.fsync(f.fileno())
        if not _journal_current(jdir, jdir_identity, journal, journal_identity):
            raise OSError("journal boundary changed during write")
    except OSError as e:
        # No journal -> no transactional guarantee. With a prior live, keep last-good and
        # discard the unlanded refresh. With NO prior (first promotion), attempt the rename
        # directly (nothing to lose) and RETAIN the staging on failure so a re-export
        # recovers it — never delete the only completed copy (P2-B06).
        log.warning("store promotion for %s: could not write journal (%s: %s)",
                    live.name, type(e).__name__, e)
        if fd is not None:
            try:
                os.close(fd)
            except OSError:  # silent-ok: best-effort close after a rejected create
                pass
            fd = None
        if (created_journal and journal_identity is not None
                and _allowed("failed-journal cleanup")):
            _unlink_bound_journal(jdir, jdir_identity,
                                  journal, journal_identity)
        if had_prior:
            if (_plain_entry_current(staged, staged_identity, directory=True)
                    and _allowed("unlanded-staging cleanup")):
                _rmtree(staged, staged_identity)
            return False
        if not _allowed("unjournaled first promotion"):
            return False
        if (not _plain_entry_current(staged, staged_identity, directory=True)
                or not _entry_absent_lstat(live)):
            return False
        try:
            staged.rename(live)
            return (_plain_entry_current(live, staged_identity, directory=True)
                    and _allowed("unjournaled promotion completion"))
        except OSError:
            log.warning("first promotion for %s: rename failed and no journal — staging "
                        "RETAINED for re-export", live.name)
            return False

    if (journal_identity is None
            or not _journal_current(jdir, jdir_identity,
                                    journal, journal_identity)):
        log.error("store promotion for %s: journal boundary changed after write; "
                  "journal and copies retained", live.name)
        return False

    def transaction_guard(action="promotion cleanup"):
        if not _allowed(action):
            return False
        if not _plain_entry_current(live.parent, parent_identity, directory=True):
            log.warning("store promotion for %s: destination parent changed before %s; "
                        "journal and copies retained", live.name, action)
            return False
        if not _journal_current(jdir, jdir_identity, journal, journal_identity):
            log.warning("store promotion for %s: journal directory/file changed before %s; "
                        "journal and copies retained", live.name, action)
            return False
        return True

    def mutation_guard():
        return transaction_guard("promotion cleanup")

    def unlink_journal():
        return _unlink_bound_journal(jdir, jdir_identity,
                                     journal, journal_identity)

    if not had_prior:
        # First promotion: no backup to make. On a rename failure RETAIN the staging + the
        # journal so recover_promotions promotes the surviving staging next launch — the
        # only completed copy is never deleted (P2-B06).
        if (not transaction_guard("first-promotion rename")
                or not _plain_entry_current(staged, staged_identity, directory=True)
                or not _entry_absent_lstat(live)):
            return False
        try:
            staged.rename(live)
            if not _plain_entry_current(live, staged_identity, directory=True):
                log.error("first promotion for %s: landed directory identity changed; "
                          "journal retained", live.name)
                return False
            if not transaction_guard("first-promotion journal cleanup"):
                return False
            unlink_journal()                         # no residue (staging consumed, no backup)
            return _allowed("first-promotion completion")
        except OSError as e:
            log.warning("first promotion for %s failed (%s: %s); staging + journal RETAINED "
                        "for next-launch recovery", live.name, type(e).__name__, e)
            return False

    if (not transaction_guard("live-to-backup rename")
            or not _plain_entry_current(live, live_identity, directory=True)
            or not _plain_entry_current(staged, staged_identity, directory=True)
            or not _entry_absent_lstat(backup)):
        return False
    try:
        live.rename(backup)                          # live -> backup
    except OSError as e:
        log.warning("store promotion for %s: could not move live aside (%s: %s) — kept last-good",
                    live.name, type(e).__name__, e)
        if (transaction_guard("failed-promotion journal cleanup")
                and _plain_entry_current(live, live_identity, directory=True)):
            unlink_journal()
        if (_plain_entry_current(staged, staged_identity, directory=True)
                and _allowed("failed-promotion staging cleanup")):
            _rmtree(staged, staged_identity)
        return False
    if (not _plain_entry_current(backup, live_identity, directory=True)
            or not _entry_absent_lstat(live)
            or not transaction_guard("staging-to-live rename")
            or not _plain_entry_current(staged, staged_identity, directory=True)):
        return False
    try:
        staged.rename(live)                          # staging -> live
    except OSError as e:
        # Death window: live is now `backup`. Restore it so we never leave zero copies.
        log.warning("store promotion for %s: could not move staging into place (%s: %s) — "
                    "restoring last-good", live.name, type(e).__name__, e)
        if (not transaction_guard("last-good restore")
                or not _plain_entry_current(backup, live_identity, directory=True)
                or not _entry_absent_lstat(live)):
            return False
        try:
            backup.rename(live)                      # restore: live present again
            if not _plain_entry_current(live, live_identity, directory=True):
                log.error("store promotion for %s: restored live identity changed; "
                          "journal retained", live.name)
                return False
            # P2-R05: drop the journal only once the residual staging is confirmed gone — a
            # locked staging RETAINS the journal so recovery retries the cleanup.
            _finalize_journal(
                journal, staged, guard=mutation_guard,
                residue_identities={staged: staged_identity},
                unlink_journal=unlink_journal)
        except OSError:
            # P2-B02: restore FAILED — live missing, backup present. RETAIN the journal +
            # backup (and staging) so the next launch's recover_promotions can restore.
            # Deleting the journal here would strand the store with zero canonical copies.
            log.error("store promotion for %s: restore from backup FAILED; journal + backup "
                      "RETAINED for next-launch recovery", live.name)
        return False
    # Commit point passed. P2-R05: drop the journal only once the backup is confirmed gone —
    # a locked backup RETAINS the journal so the next launch's recovery removes the residue.
    if not _plain_entry_current(live, staged_identity, directory=True):
        log.error("store promotion for %s: promoted live identity changed; journal retained",
                  live.name)
        return False
    _finalize_journal(
        journal, backup, guard=mutation_guard,
        residue_identities={backup: live_identity},
        unlink_journal=unlink_journal)
    return (_plain_entry_current(live, staged_identity, directory=True)
            and _allowed("promotion completion"))


def _trusted_journal_names(j):
    """The (target, backup, staging) basenames from a journal record IFF it is in the
    expected promotion shape — else None (P2-B04: a malformed/planted journal must never
    let recovery touch a path outside its own store). Each name must be a single basename
    (no separators / ``..`` / absolute parts) and ``backup``/``staging`` must be exactly
    ``<target>.bak-<token>`` / ``<target>.staging``."""
    if not isinstance(j, dict):
        return None
    target, backup, staging, token = (j.get("target"), j.get("backup"),
                                      j.get("staging"), j.get("token"))
    if not all(isinstance(s, str) and s for s in (target, backup, staging, token)):
        return None
    for n in (target, backup, staging):
        if n in (".", "..") or n != Path(n).name:    # reject separators / .. / absolute
            return None
    if backup != f"{target}{_BAK_INFIX}{token}" or staging != f"{target}{_STAGING_SUFFIX}":
        return None
    return target, backup, staging


def _journal_gen(journal, jdir=None, jdir_identity=None, journal_identity=None):
    """The transaction GENERATION recorded in a journal (P2-B07), or 0 when absent / unreadable /
    non-int. Higher = newer; recovery processes a target's journals highest-`gen` first."""
    journal = Path(journal)
    jdir = Path(jdir) if jdir is not None else journal.parent
    if jdir_identity is None:
        jdir_identity = _plain_journal_dir_identity(jdir)
    if journal_identity is None:
        journal_identity = _plain_journal_identity(jdir, jdir_identity, journal)
    if journal_identity is None:
        return 0
    value = _read_bound_journal(
        jdir, jdir_identity, journal, journal_identity)
    if not isinstance(value, dict):
        return 0
    g = value.get("gen")
    return g if isinstance(g, int) and not isinstance(g, bool) else 0


def _next_generation(jdir, target_name, jdir_identity=None):
    """The generation to stamp on a NEW promotion of `target_name`: strictly greater than EVERY
    valid existing same-target journal's generation (P2-B07). DURABLE + monotonic — derived from
    the on-disk journals under the single-writer promotion boundary, NOT wall-clock, so it cannot
    regress across application restarts or system-clock corrections (an older retained cleanup
    journal can never out-rank a later promotion). A first promotion (no prior same-target
    journal) is generation 1."""
    highest = 0
    jdir = Path(jdir)
    if jdir_identity is None:
        jdir_identity = _plain_journal_dir_identity(jdir)
        if jdir_identity is None:
            try:
                jdir.lstat()
            except OSError:
                return 1                             # no .promote dir yet
            return None                              # existing but unsafe boundary
    if not _journal_dir_current(jdir, jdir_identity):
        return None
    try:
        journals = [p for p in jdir.iterdir() if p.suffix == ".json"]
    except OSError:
        return None
    for journal in journals:
        journal_identity = _plain_journal_identity(
            jdir, jdir_identity, journal)
        if journal_identity is None:
            continue                                 # never follow a journal symlink/reparse
        j = _read_bound_journal(
            jdir, jdir_identity, journal, journal_identity)
        if j is _JOURNAL_UNSAFE:
            return None                              # directory/file changed before/during read
        if j is _JOURNAL_INVALID:
            continue                                 # junk -> recovery drops it; ignore for gen
        names = _trusted_journal_names(j)
        if names is None or names[0] != target_name:
            continue                                 # malformed / a DIFFERENT target — irrelevant
        g = j.get("gen")
        if isinstance(g, int) and not isinstance(g, bool) and g > highest:
            highest = g
    return highest + 1 if _journal_dir_current(jdir, jdir_identity) else None


def _recover_one(jdir, journal, is_owned, jdir_identity=None,
                 journal_identity=None):
    """Act on ONE promotion journal in an already-OWNERSHIP-CONFIRMED location (the caller
    `recover_promotions` proves `is_owned(store_root, None)` for `jdir.parent` BEFORE this is
    reached. Every mutation rechecks that same caller-owned lease so a different directory
    moved onto the pathname cannot inherit recovery authority — P2-B04 / CMP-AUD-090).
    The journal is dropped ONLY after a canonical copy is PROVEN usable — a directory holding
    a report artifact, NOT a bare ``exists()`` (P2-B02) — AND every journal-owned residue is
    confirmed GONE (P2-R05). When `live` is absent / an invalid placeholder, a valid backup
    (then a valid staging) is restored; an empty/disposable placeholder is displaced first,
    but FOREIGN content is a conflict that retains everything; a blocked restore likewise."""
    jdir, journal = Path(jdir), Path(journal)
    parent = jdir.parent
    jdir_identity = (jdir_identity
                     if jdir_identity is not None
                     else _plain_journal_dir_identity(jdir))
    journal_identity = (journal_identity
                        if journal_identity is not None
                        else _plain_journal_identity(jdir, jdir_identity, journal))

    def _guard(target_name, action):
        if (jdir_identity is None or journal_identity is None
                or not _journal_current(jdir, jdir_identity,
                                        journal, journal_identity)):
            log.warning("promotion recovery: journal directory/file changed before %s; "
                        "external path retained", action)
            return False
        try:
            current = bool(is_owned(parent, target_name))
        except Exception as e:  # noqa: BLE001 — recovery remains best-effort
            log.warning("promotion recovery: ownership check failed before %s "
                        "(%s: %s); retained", action, type(e).__name__, e)
            return False
        if not current:
            log.warning("promotion recovery: store ownership changed before %s; "
                        "journal and copies retained", action)
        return current
    if not _guard(None, "journal read"):
        return
    j = _read_bound_journal(
        jdir, jdir_identity, journal, journal_identity)
    if j is _JOURNAL_UNSAFE:
        return                                      # replacement/reparse: touch nothing
    if j is _JOURNAL_INVALID:
        if _guard(None, "unreadable-journal cleanup"):
            _unlink_bound_journal(jdir, jdir_identity,
                                  journal, journal_identity)
        return
    names = _trusted_journal_names(j)
    if names is None:
        log.warning("promotion recovery: dropping a shape-invalid journal %s in an owned store "
                    "(no path touched)", journal.name)
        if _guard(None, "invalid-journal cleanup"):
            _unlink_bound_journal(jdir, jdir_identity,
                                  journal, journal_identity)
        return
    if not _guard(names[0], "journal recovery"):    # P2-B04: target + leased root
        log.warning("promotion recovery: ignoring journal %s — target %r is not a known report "
                    "(no path touched)", journal.name, names[0])
        return
    target, backup, staging = (parent / n for n in names)

    def target_guard():
        return _guard(names[0], "recovery mutation")

    def unlink_journal():
        return _unlink_bound_journal(jdir, jdir_identity,
                                     journal, journal_identity)

    def residue_identities():
        """Bind ordinary residues at the final cleanup decision."""
        return {
            backup: (_plain_entry_identity(backup, directory=True)
                     if _entry_present_lstat(backup) else _ABSENT_ENTRY),
            staging: (_plain_entry_identity(staging, directory=True)
                      if _entry_present_lstat(staging) else _ABSENT_ENTRY),
        }

    target_identity = _plain_entry_identity(target, directory=True)
    if (target_identity is not None and is_usable_store(target)
            and _plain_entry_current(target, target_identity, directory=True)):
        _finalize_journal(
            journal, backup, staging, guard=target_guard,
            residue_identities=residue_identities(),
            unlink_journal=unlink_journal)
        return
    # `target` is absent OR an invalid placeholder. Restore from a valid backup, else a
    # valid surviving staging (the first-promotion case). Validate the source IS a real
    # store so we never restore garbage over nothing.
    for src_name, src in (("backup", backup), ("staging", staging)):
        src_identity = _plain_entry_identity(src, directory=True)
        if (src_identity is None or not is_usable_store(src)
                or not _plain_entry_current(src, src_identity, directory=True)):
            continue
        target_present = _entry_present_lstat(target)
        target_identity = (_plain_entry_identity(target, directory=True)
                           if target_present else None)
        if target_present and (target_identity is None
                               or not _dir_is_disposable_placeholder(target)):
            log.error("promotion recovery: %s exists with foreign content — cannot restore "
                      "from %s; journal + copies RETAINED for manual resolution",
                      target.name, src_name)
            return                                   # genuine conflict — touch nothing
        if target_present:
            if not target_guard():
                return
            if not _rmtree_gone(target, target_identity):
                return                               # changed/locked placeholder: retain all
        try:
            if (not target_guard()
                    or not _entry_absent_lstat(target)
                    or not _plain_entry_current(src, src_identity, directory=True)):
                return
            src.rename(target)
            if not _plain_entry_current(target, src_identity, directory=True):
                log.error("promotion recovery: restored %s identity changed; retained",
                          target.name)
                return
            log.info("promotion recovery: restored %s from %s", target.name, src_name)
        except OSError as e:                         # blocked — RETAIN journal + copies for retry
            log.error("promotion recovery: could not restore %s from %s (%s: %s); retained",
                      target.name, src_name, type(e).__name__, e)
            return
        _finalize_journal(
            journal, backup, staging, guard=target_guard,
            residue_identities=residue_identities(),
            unlink_journal=unlink_journal)
        return
    # No usable backup or staging. Prefer RETAINING any residue (harmless, and a human can
    # inspect it) over deleting an unvalidated copy; drop the journal only when nothing is left.
    if _entry_present_lstat(backup) or _entry_present_lstat(staging):
        log.error("promotion recovery: no usable copy for %s; residue RETAINED for diagnosis",
                  target.name)
        return
    if target_guard():
        unlink_journal()                             # nothing recoverable: drop the journal


def recover_promotions(root, is_owned):
    """Startup recovery sweep (F2): repair any interrupted store promotion under `root`.

    `is_owned(store_root: Path, target_name | None) -> bool` is the caller-supplied OWNERSHIP
    context (P2-B04). Ownership is established from the COMPLETE LOCATION BEFORE any journal is
    read or deleted, then revalidated before each mutation: for each discovered
    ``<store_root>/.promote`` the LOCATION gate `is_owned(store_root, None)` must pass. The
    updater requires a direct known ``<src>-<env>`` child plus a current purpose-bound marker
    under user destinations; its exact app-private OUTPUT_ROOT keeps legacy name compatibility.
    Otherwise the whole ``.promote`` (including any malformed journals in it) is left
    UNTOUCHED. So a nested,
    valid-NAME-but-wrong-LOCATION tree (e.g. ``<root>/UnrelatedProject/ssor-prod/.promote``)
    is never acted on. Inside an owned location, each shape-valid journal's TARGET is then
    validated via `is_owned(store_root, target)`.

    Journals are processed NEWEST GENERATION FIRST (by the durable monotonic ``gen``; P2-B07), so
    if two journals name the same target — an older retained-after-locked-cleanup one plus a newer
    interrupted one — the newer generation restores/keeps the true last-good first and the older
    journal can only become residue-cleanup against an already-canonical live; an older generation
    can never roll back over or delete a newer one, regardless of filesystem order OR a wall-clock
    regression (``gen`` is derived from existing journals, not the clock). There is NO journal-free
    orphan sweep. Idempotent + best-effort; never raises."""
    root = Path(root)
    root_identity = _plain_entry_identity(root, directory=True)
    if root_identity is None:
        return
    # Legitimate stores are either ``root`` itself (mechanics/legacy recovery)
    # or direct children of the configured destination. A recursive glob can
    # traverse a directory junction before the later `.promote` lstat check;
    # enumerate only ordinary direct children and never descend through reparses.
    try:
        root_entries = list(root.iterdir())
    except OSError:
        return
    if not _plain_entry_current(root, root_identity, directory=True):
        return
    promote_dirs = [root / _PROMOTE_DIRNAME]
    promote_dirs.extend(
        entry / _PROMOTE_DIRNAME
        for entry in root_entries
        if _plain_entry_identity(entry, directory=True) is not None)
    for jdir in promote_dirs:
        if not _entry_present_lstat(jdir):
            continue
        jdir_identity = _plain_journal_dir_identity(jdir)
        if jdir_identity is None:
            log.warning("promotion recovery: refusing unsafe/reparse .promote at %s "
                        "(no journal read or touched)", jdir)
            continue
        if not is_owned(jdir.parent, None):          # P2-B04: LOCATION gate BEFORE any read/delete
            log.warning("promotion recovery: skipping a non-app-owned .promote at %s (no journal "
                        "read or touched)", jdir.parent)
            continue
        try:
            entries = list(jdir.iterdir())
        except OSError:
            continue
        if not _journal_dir_current(jdir, jdir_identity):
            continue
        journals = []
        for p in entries:
            if p.suffix != ".json":
                continue
            identity = _plain_journal_identity(jdir, jdir_identity, p)
            if identity is not None:
                journals.append((p, identity))
        # P2-B07: highest generation first (stable tie-break on name) — order-independent of the
        # filesystem listing AND of wall-clock, so an older same-target journal never restores
        # before the newer one even after a system-clock regression.
        ordered = sorted(
            journals,
            key=lambda item: (_journal_gen(item[0], jdir, jdir_identity, item[1]),
                              item[0].name),
            reverse=True)
        for journal, journal_identity in ordered:
            _recover_one(jdir, journal, is_owned, jdir_identity, journal_identity)
        try:                                         # drop the .promote dir if now empty
            if (is_owned(jdir.parent, None)
                    and _journal_dir_current(jdir, jdir_identity)):
                jdir.rmdir()
        except OSError:
            pass
