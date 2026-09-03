"""Build the TSMIS-vs-TSN Highway Sequence (Highway Locations) comparison
workbook (FLAT, keyed on route + county + postmile).

Both sides are highway postmile-sequence listings — the direct analog of the
Highway Log comparison, with the same "TSN lists more segment breaks, TSMIS more
realignment markers" one-sided behavior. Reconciled by hand against the 6.19
ground truth (docs/tsn-parsers.md):

  * TSMIS side — the CONSOLIDATED Highway Sequence workbook (sheet "Highway
    Locations", leading Route column). Its header has two unnamed columns (a
    postmile prefix and an equate suffix), so it is read BY POSITION:
      0 Route · 1 County · 2 City · 3 prefix · 4 PM · 5 suffix · 6 HG · 7 FT
      · 8 Distance To Next Point · 9 Description
    The canonical postmile re-glues prefix+PM+suffix ("R" + "000.129" -> "R000.129").
  * TSN side — the normalized workbook built by consolidate_tsn_highway_sequence
    from the district PDFs (its NORMALIZED_HEADER, postmile already glued).

CALIFORNIA postmiles are COUNTY-RELATIVE (a route restarts at 000.000 in each
county), so route+PM is NOT unique across a route — the key is route + county +
postmile (county folded into the key via key_normalizer, kept as its own visible
column). Landmarks that still share a (route, county, PM) pair (e.g. a "COUNTY
BEGIN" marker at the same postmile) are matched by data similarity in compare_core.

Console-free; engine in compare_core.
"""
from collections import Counter
from dataclasses import replace
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import compare_tsn_common as ctc
from compare_tsn_common import (load_consolidated_rows, row_has_data,
                                suggest_route_name)
import comparison_contract as cc
import consolidate_tsn_highway_sequence as tsn_hsl
import consolidation_meta
from compare_core import CompareSchema, normalize_value

REPORT_NAME = "Highway Sequence"
TSMIS_SHEET = "Highway Locations"          # consolidated sheet (Route prepended)

# CMP-AUD-034: the EXACT consolidated-TSMIS header (['Route'] + the export's own
# layout, its two header-less columns as ''). _tsmis_row reads by position, so this
# is bound exactly — a shifted/relabelled/wrong-edition header is refused, never
# mis-mapped. Verified statewide-stable + data-source/env/edition-independent
# (2026-07-17 census, identical across all six source/env combos); the Highway
# Sequence (PDF) consolidator emits the IDENTICAL header, so the PDF-vs-TSN /
# PDF-vs-Excel flavors that also load through _load_tsmis stay valid.
_TSMIS_HEADER = ["Route", "County", "City", "", "PM", "", "HG", "FT",
                 "Distance To Next Point", "Description"]

KEY = "PM"
SHARED_HEADER = ["County", "PM", "City", "HG", "FT",
                 "Distance To Next Point", "Description"]
KEY_FIELD = SHARED_HEADER.index(KEY)       # 1

# EVERY column is compared (owner decision 2026-08-10). Three of them used to be
# context because each is noisy for a STRUCTURAL reason rather than a data
# disagreement — that reasoning is still true and is now disclosed on the Notes
# sheet instead of used to suppress the count:
#   HG       — TSMIS leaves the highway-group blank for whole counties while TSN
#              always fills it (U/D), so expect many blank-vs-U cells.
#   City     — TSN assigns a city code far more aggressively than TSMIS (TSN tags
#              the nearest incorporated place; TSMIS only within strict limits).
#   Distance — "distance to next point" is measured to each system's OWN next
#              listed point; TSN lists more (finer) breaks, so its gap is usually
#              smaller (TSMIS 003.572 vs TSN 000.174 at the same postmile) — an
#              artifact of listing granularity.
# County is part of the key (always equal within a matched pair); PM is the key.
CONTEXT_FIELDS = ()

# Consolidated-TSMIS VALUE positions (Route at 0; verified on the 6.19 set).
_TSMIS = {"route": 0, "county": 1, "city": 2, "prefix": 3, "pm": 4,
          "suffix": 5, "hg": 6, "ft": 7, "dist": 8, "desc": 9}


# --------------------------------------------------------------------------- #
# normalization
# --------------------------------------------------------------------------- #
_WS_RE = re.compile(r"[\t\n\r\f\v]")


def _v(x):
    """compare_core.normalize_value plus the OOXML `_xHHHH_` escape decode and
    tab/newline collapse. The TSMIS export pads Description with trailing tabs
    that Excel TRIM does not strip, and its censused `_x000d_` literals are
    encoded CRs (CMP-AUD-197): the Stage-8 oracle xlsx-unescapes every side it
    reads, and the decode is byte-equivalent to openpyxl's — a no-op on the
    escape-free raw-TSN and PDF-render sides."""
    nv = normalize_value(x)
    if not isinstance(nv, str):
        return nv
    return _WS_RE.sub(" ", ctc.decode_ooxml_escapes(nv)).strip()


def _raw_text(x):
    """A cell's lossless source text for identity claims."""
    return "" if x is None else str(x)


def _norm_county(x):
    """Canonical county code. TSMIS writes a trailing period on several codes
    ("LA.", "SB.", "SM.", "SF.", "CC.", "DN.", "ED.", "SD.", "SJ.") that the TSN
    report omits — strip it so the county-relative postmile key matches."""
    s = "" if x is None else str(x).strip().rstrip(".")
    return s.upper()


_DESC_PREFIX_RE = re.compile(r"^(\d{1,3}[A-Z]?)/")
_ROUTE_TOKEN_RE = re.compile(r"(\d{1,3})([A-Za-z]?)")


def _canon_route(tok):
    """A route token in canonical form ("1" -> "001", "14u" -> "014U"), or None
    for a non-route token."""
    m = _ROUTE_TOKEN_RE.fullmatch(("" if tok is None else str(tok)).strip())
    return (m.group(1).zfill(3) + m.group(2).upper()) if m else None


def _desc_plain(x):
    """Description with whitespace runs collapsed to one space — and NOTHING
    else. This is the TSN side (CMP-AUD-204: raw TSN keeps 154 numeric-prefix
    Descriptions — 46 of them deliberately naming a DIFFERENT route — and every
    one is an authoritative source claim; deleting them false-cleaned 81 real
    differences per leg) and both sides of the PDF-vs-Excel self-check."""
    s = _v(x)
    return re.sub(r"\s+", " ", s).strip() if isinstance(s, str) else s


def _desc_tsmis(x, route):
    """The TSMIS Description for the vs-TSN legs: whitespace-collapsed, with the
    separately added leading route label ("001/NB OFF TO DOHENY PK RD") removed
    ONLY when the token names this row's OWN route (CMP-AUD-204). A leading
    cross-route or nested token is genuine source text — TSN prints it too, and
    a pattern-blind strip both hid real differences and fabricated others."""
    s = _desc_plain(x)
    if not isinstance(s, str):
        return s
    m = _DESC_PREFIX_RE.match(s)
    if m and _canon_route(m.group(1)) == _canon_route(route):
        return s[m.end():].lstrip()
    return s


def _glue_pm(prefix, pm, suffix):
    """Re-glue the TSMIS prefix/PM/suffix columns into the canonical postmile the
    TSN side prints ("R" + "000.129" + "" -> "R000.129"; "" + "050.025" + "E"
    -> "050.025E")."""
    p = "" if prefix is None else str(prefix).strip()
    m = "" if pm is None else str(pm).strip()
    s = "" if suffix is None else str(suffix).strip()
    return f"{p}{m}{s}"


# --------------------------------------------------------------------------- #
# physical identity (CMP-AUD-045/199)
# --------------------------------------------------------------------------- #
# The approved HSL identity is (Route, County, complete glued postmile):
# California postmiles are county-relative, and HSL's canonical postmile keeps
# its zero padding, realignment prefix, and — on the vs-TSN and cross-env paths
# — the equate suffix exactly as printed ("R001.000E"). Both sources
# legitimately print rows with NO county (46 raw TSN equate annotations precede
# any county context, CMP-AUD-158) or NO postmile (five TSMIS rows per render),
# so those key under explicit reserved tokens that can never collide with a
# real county/postmile: unknown ownership is disclosed in the key column, never
# dropped, backfilled, or crashed on.
_NO_COUNTY_KEY = "(county not printed)"
_NO_PM_KEY = "(no postmile printed)"


def _physical_pm_key(route, county_raw, pm_glued, claims, source_hint):
    """The HSL PhysicalKey: canonical (route, county, glued postmile) identity
    with the raw prefix/PM/suffix (or the printed glued token) conserved as
    lossless claims. `source_hint` names the failing side in errors."""
    county = _norm_county(county_raw) or _NO_COUNTY_KEY
    pm = pm_glued or _NO_PM_KEY
    try:
        identity = cc.make_physical_identity(
            route, county, pm,
            tuple(cc.RawIdentityClaim(name, value) for name, value in claims),
            f"{route} / {county} / {pm}")
        return cc.physical_key(pm, identity)
    except ValueError as e:
        raise ValueError(f"{source_hint}: could not build the physical "
                         f"identity for route {route!r} PM {pm!r}: {e}")


# --------------------------------------------------------------------------- #
# loaders -> consolidated-shape rows ([route, *SHARED_HEADER])
# --------------------------------------------------------------------------- #
def _tsmis_row(r):
    def at(i):
        return r[i] if i < len(r) else None
    route = _v(at(_TSMIS["route"]))
    county_raw = at(_TSMIS["county"])
    prefix, pm, suffix = at(_TSMIS["prefix"]), at(_TSMIS["pm"]), at(_TSMIS["suffix"])
    key = _physical_pm_key(
        route, county_raw, _glue_pm(prefix, pm, suffix),
        (("route", _raw_text(at(_TSMIS["route"]))),
         ("county", _raw_text(county_raw)),
         ("postmile_prefix", _raw_text(prefix)),
         ("postmile", _raw_text(pm)),
         ("postmile_suffix", _raw_text(suffix))),
        "the consolidated TSMIS workbook")
    return [route,
            _norm_county(county_raw),
            key,
            _v(at(_TSMIS["city"])),
            _v(at(_TSMIS["hg"])),
            _v(at(_TSMIS["ft"])),
            _v(at(_TSMIS["dist"])),
            _desc_tsmis(at(_TSMIS["desc"]), route)]


def _load_tsmis_raw(path):
    """The consolidated workbook's data rows exactly as exported (the header gate
    applied, nothing normalized) — the shape `seat_equate_suffixes` rewrites."""
    return load_consolidated_rows(
        path, TSMIS_SHEET,
        missing_sheet_hint="pick the consolidated TSMIS Highway Sequence workbook.",
        bad_header_msg="isn't a CONSOLIDATED Highway Sequence workbook in the "
                       "current site layout (expected a leading 'Route' column and "
                       "the exact export header) — consolidate a fresh export first.",
        header_ok=ctc.exact_consolidated_header_ok(_TSMIS_HEADER),  # CMP-AUD-034
        row_transform=list)


def _load_tsmis(path):
    """The consolidated TSMIS rows in comparison shape, exactly as exported.
    The Excel-vs-TSN leg seats the export's equate suffix first (`_load_pair`,
    the equate-seat section below); the PDF-edition flavors load through this
    directly because the print already seats it the TSN way."""
    rows, has_route = _load_tsmis_raw(path)
    return [_tsmis_row(r) for r in rows], has_route


def _normalization_version(wb):
    """The normalized workbook's declared version (0 for a pre-v4 workbook —
    the rows sheet kept its SHAPE across v4, so the marker sheet is the only
    reliable signal on a bare file; the library path additionally auto-rebuilds
    via report_catalog's normalization_version, D2)."""
    if tsn_hsl.MARKER_SHEET not in wb.sheetnames:
        return 0
    for r in wb[tsn_hsl.MARKER_SHEET].iter_rows(values_only=True):
        if r and str(r[0]).strip() == "Normalization version":
            try:
                return int(r[1])
            except (TypeError, ValueError, IndexError):  # silent-ok: a malformed marker reads as version 0 — the caller then refuses with the rebuild hint (fail-safe)
                return 0
    return 0


def _load_tsn(path):
    """The normalized TSN workbook (consolidate_tsn_highway_sequence output):
    NORMALIZED_HEADER = [Route] + SHARED_HEADER, read positionally. Refuses a
    pre-v4 workbook: it is missing the 46 blank-County equate annotations and
    the printed pointer tokens, and carries an invented join comma
    (CMP-AUD-156/158/159) — silently comparing it would resurrect all three."""
    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        sn = tsn_hsl.NORMALIZED_SHEET
        if sn not in wb.sheetnames:
            raise ValueError(f"{name} has no '{sn}' sheet — pick the normalized TSN "
                             "Highway Sequence workbook (built from the district PDFs).")
        if _normalization_version(wb) < tsn_hsl.NORMALIZATION_VERSION:
            raise ValueError(
                f"{name} was built by an older TSN Highway Sequence converter "
                "(pre-v4: pointer tokens blanked, pre-county equates dropped, "
                "an invented join comma) — rebuild the TSN library and pick "
                "the fresh normalized workbook.")
        it = wb[sn].iter_rows(values_only=True)
        header = [("" if c is None else str(c).strip())
                  for c in (next(it, None) or ())]
        # CMP-AUD-033: bind the header to the exact ["Route"] + SHARED_HEADER
        # layout (no sidecars) before reading BY POSITION — the loader trusted
        # any sheet, so a reordered/renamed header silently mis-mapped columns.
        ctc.require_shared_header_prefix(
            header, ["Route"] + SHARED_HEADER, (), name, REPORT_NAME)
        width = len(SHARED_HEADER) + 1         # Route + fields
        pm_idx = 1 + KEY_FIELD
        county_idx = 1 + SHARED_HEADER.index("County")
        desc_idx = 1 + SHARED_HEADER.index("Description")
        rows = []
        for r in it:
            if not row_has_data(r):
                continue
            r = list(r)[:width] + [None] * max(0, width - len(r))
            route = _v(r[0])
            key = _physical_pm_key(
                route, r[county_idx],
                "" if r[pm_idx] is None else str(r[pm_idx]).strip(),
                (("route", _raw_text(r[0])),
                 ("county", _raw_text(r[county_idx])),
                 ("postmile", _raw_text(r[pm_idx]))),
                "the normalized TSN workbook")
            row = [route if i == 0 else key if i == pm_idx
                   else _norm_county(r[i]) if i == county_idx
                   else _desc_plain(r[i]) if i == desc_idx
                   else _v(r[i])
                   for i in range(width)]
            rows.append(row)
        return rows, True
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# the equate seat — TSN declares the relation, the Excel export's suffix is
# seated where TSN and the print put it (roadmap E10)
# --------------------------------------------------------------------------- #
# A postmile EQUATION is one fact the three renders spell differently by
# design. TSN and the TSMIS print both write an annotation line at the
# realignment postmile ("EQUATES TO", flags blank) and put the "E" suffix on
# the equated postmile's OWN row. The Excel export has no annotation
# convention: it folds the marker onto the realignment record and, about a
# quarter of the time, seats the "E" THERE instead of on the target. The
# vs-TSN identity is the complete glued postmile (CMP-AUD-045), so a relation
# the export seats on the other member missed on BOTH rows — TSN "R004.629" /
# Excel "R004.629E" at the annotation, TSN "R004.508E" / Excel "R004.508" at
# the target — and the comparison reported a location as present in one
# system only when it is in both.
#
# Measured on the 2026-08-31 statewide pull (raw census, before any rule): the
# PDF edition paired 444 TSN keys the Excel edition missed, and 444 of the 444
# were this class — 214 annotation rows, 229 E-suffixed targets, and one data
# row at an annotation postmile the export had seated an "E" on. TSN declares
# 998 relations statewide; every one is a bare-postmile annotation followed by
# exactly one "E" row of the same route, and no "E" row is unclaimed.
#
# The rule mirrors the PDF-vs-Excel self check (compare_highway_sequence_pdf,
# HF-06 / PCOA-FINAL-011) with TSN in the print's role: TSN DECLARES each
# relation; the export's suffix is moved from the annotation's postmile to the
# relation's target row, and nothing else changes — the marker text, the
# annotation row's HG/FT and every other cell are compared exactly as before,
# so the by-design Description/FT differences at a paired equate stay counted
# and stay disclosed in the Notes. It FAILS OPEN: a relation the export does
# not carry, whose rows it cannot single out, or whose two rows both carry a
# suffix is left exactly as exported. Refusing can only leave a one-sided row
# visible; it can never invent a pair.
_EQUATE_MARKER = "EQUATES TO"
_GLUED_PM_RE = re.compile(r"^(?P<bare>[A-Z]?\d{3}\.\d{3})(?P<suffix>[A-Z]?)$")


def _split_glued(pm):
    """A glued postmile -> (bare postmile, equate suffix), or None when the
    token is not a printed postmile (the reserved no-postmile marker)."""
    m = _GLUED_PM_RE.match("" if pm is None else str(pm).strip())
    return (m.group("bare"), m.group("suffix")) if m else None


def equate_relations(rows_n):
    """The equate relations the LOADED TSN rows declare, in file order, each
    (route, annotation county, annotation bare PM, target county, target bare
    PM, target suffix, target Description).

    An annotation is a bare-postmile "EQUATES TO" row; its target is the next
    row of the same route carrying a suffix, searched forward and stopping at
    the next annotation (statewide it is always the very next row). An
    annotation with no such target, or printed with a suffix of its own,
    declares nothing."""
    pm_idx = 1 + KEY_FIELD
    county_idx = 1 + SHARED_HEADER.index("County")
    desc_idx = 1 + SHARED_HEADER.index("Description")
    out = []
    for i, row in enumerate(rows_n):
        if row[desc_idx] != _EQUATE_MARKER:
            continue
        split = _split_glued(row[pm_idx])
        if split is None or split[1]:
            continue
        target = None
        for j in range(i + 1, len(rows_n)):
            nxt = rows_n[j]
            if nxt[0] != row[0] or nxt[desc_idx] == _EQUATE_MARKER:
                break
            t_split = _split_glued(nxt[pm_idx])
            if t_split is not None and t_split[1]:
                target = (nxt[county_idx], t_split[0], t_split[1], nxt[desc_idx])
                break
        if target is not None:
            out.append((row[0], row[county_idx], split[0]) + target)
    return tuple(out)


def _raw_cell(row, i):
    return "" if i >= len(row) or row[i] is None else str(row[i]).strip()


def _with_suffix(row, suffix):
    """A copy of one raw export row with its suffix cell set (None = blank,
    the export's own representation)."""
    out = list(row) + [None] * max(0, _TSMIS["suffix"] + 1 - len(row))
    out[_TSMIS["suffix"]] = suffix or None
    return out


def seat_equate_suffixes(raw_rows, relations, stats=None):
    """A NEW list of the export's raw rows with each TSN-declared relation's
    suffix seated on its target row. `stats` (an optional dict) receives the
    counts: `declared`, `seated`, and why every other relation was left as
    exported. Rows are resolved against the ORIGINAL rows, never the partly
    rewritten copy, so one relation's move can never shift the next one's
    view of the export."""
    P = _TSMIS
    groups = {}
    for i, r in enumerate(raw_rows):
        route = _v(r[P["route"]] if len(r) > P["route"] else None)
        bare = _glue_pm(_raw_cell(r, P["prefix"]), _raw_cell(r, P["pm"]), None)
        groups.setdefault((route, _norm_county(_raw_cell(r, P["county"])), bare),
                          []).append(i)
    counts = Counter()
    out = [list(r) for r in raw_rows]
    for route, a_county, a_bare, t_county, t_bare, t_suffix, t_desc in relations:
        counts["declared"] += 1
        ann_group = groups.get((route, a_county, a_bare), ())
        if not ann_group:
            # the 46 pre-county annotations (CMP-AUD-158) and the breaks the
            # export does not list at all — nothing here to seat anything on
            counts["realignment record not exported"] += 1
            continue
        carriers = [i for i in ann_group if _raw_cell(raw_rows[i], P["suffix"])]
        if not carriers:
            # already on the target, or the export carries none anywhere —
            # either way nothing to move, and a missing marker stays visible
            counts["no suffix at the realignment record"] += 1
            continue
        if len(carriers) > 1 or _raw_cell(raw_rows[carriers[0]], P["suffix"]) != t_suffix:
            counts["ambiguous"] += 1
            continue
        tgt_group = groups.get((route, t_county, t_bare), ())
        if not tgt_group:
            counts["target not exported"] += 1
            continue
        if any(_raw_cell(raw_rows[i], P["suffix"]) for i in tgt_group):
            counts["both rows carry a suffix"] += 1
            continue
        if len(tgt_group) == 1:
            tgt = tgt_group[0]
        else:
            # the export lists several rows at the target postmile: the one
            # that IS TSN's target is the one printing its Description
            matches = [i for i in tgt_group
                       if _desc_tsmis(_raw_cell(raw_rows[i], P["desc"]), route) == t_desc]
            if len(matches) != 1:
                counts["ambiguous"] += 1
                continue
            tgt = matches[0]
        ann = carriers[0]
        out[ann] = _with_suffix(out[ann], None)
        out[tgt] = _with_suffix(out[tgt], t_suffix)
        counts["seated"] += 1
    if stats is not None:
        stats.update(counts)
    return out


_SEAT_LEFT_REASONS = ("no suffix at the realignment record",
                      "realignment record not exported", "target not exported",
                      "both rows carry a suffix", "ambiguous")


def _equate_seat_disclosure(stats):
    """The run-resolved disclosure line for the Summary and the Notes."""
    def line():
        if not stats:
            return ""
        left = ", ".join(f"{stats[k]:,} {k}" for k in _SEAT_LEFT_REASONS if stats.get(k))
        return (f"Equate suffix seat: TSN declares {stats.get('declared', 0):,} equate "
                "relations (an \"EQUATES TO\" annotation plus the next row carrying "
                "the \"E\"). The Excel export seats that \"E\" on the realignment "
                "record about a quarter of the time, where TSN puts it on the "
                "equated postmile's own row; since the key is the complete glued "
                "postmile, that seat alone made BOTH rows of such a relation "
                f"one-sided. The suffix was moved to the target row for "
                f"{stats.get('seated', 0):,} relations before comparing — nothing "
                "else changed, so the annotation's Description/HG/FT are still "
                "compared and counted"
                + (f" — and left exactly as exported for {left}." if left else "."))
    return line


# --------------------------------------------------------------------------- #
# Notes sheet — the INDICATOR for the context fields + key
# --------------------------------------------------------------------------- #
_NOTES_TITLE = "Highway Sequence — TSMIS vs TSN: comparison notes"
_NOTES_LINES = (
    "Rows are keyed on Route + County + Postmile. California postmiles are "
    "county-relative (a route restarts at 000.000 in each county it crosses), so "
    "the postmile alone is not unique across a route — County is part of the key.",
    "The postmile carries a glued realignment prefix (\"R000.129\") and/or an equate "
    "suffix (\"050.025E\"); the TSMIS prefix/PM/suffix columns are re-glued to match.",
    "A handful of rows print with NO county (46 statewide TSN \"EQUATES TO\" "
    "annotations that appear before the route's first county-bearing row — TSN's own "
    "cover warns equate ownership may be wrong) or NO postmile (five TSMIS rows). "
    "They key under the explicit \"(county not printed)\" / \"(no postmile printed)\" "
    "markers and surface honestly, usually one-sided — never dropped or backfilled.",
    "One-sided rows are expected and honest: TSN lists every segment break (including "
    "unnamed ones) and prints equate points as \"EQUATES TO\" annotations, while TSMIS "
    "omits most unnamed breaks and records the equate as an \"END R REALIGNMENT\" row.",
    "Equate points that BOTH systems mark pair up by design and then differ on "
    "purpose: TSN's bare \"EQUATES TO\" annotation carries no feature type, so the "
    "pair surfaces as a Description difference (TSMIS's realignment/route-break "
    "label vs \"EQUATES TO\") and usually an FT difference (TSMIS \"H\" vs TSN "
    "blank). Nearly all FT differences statewide are this class; the few remaining "
    "are genuine feature-type disagreements (H vs I, R vs H).",
    "The Excel export seats the equate \"E\" suffix on the realignment record about a "
    "quarter of the time, where TSN (and the TSMIS print) put it on the equated "
    "postmile's own row. Because the key is the complete glued postmile, that seat "
    "alone made BOTH rows of such a relation one-sided. TSN declares each relation "
    "(its \"EQUATES TO\" annotation plus the next row carrying the suffix); the "
    "export's suffix is moved to that target row before comparing and nothing else "
    "is changed — the annotation's Description/HG/FT are still compared. A relation "
    "the export does not carry, cannot single out, or seats on both rows is left "
    "exactly as exported; the Summary states the counts.",
    "Descriptions: the TSMIS export prepends the row's own route as a label "
    "(\"001/NB OFF TO DOHENY PK RD\") — that label alone is stripped before "
    "comparing. TSN text is compared VERBATIM: TSN's numeric route prefixes "
    "(including ones naming a DIFFERENT route) are authoritative source claims, so "
    "TSMIS \"103 SEP 53-145\" vs TSN \"1/103 SEP 53-145\" is a REAL difference. A "
    "leading cross-route token on the TSMIS side is likewise kept.",
    "EVERY column is compared, including three that are noisy for a STRUCTURAL "
    "reason rather than a data disagreement (owner decision 2026-08-10; they were "
    "context columns before): HG — TSMIS leaves the highway-group blank for whole "
    "counties while TSN always fills it, so expect many blank-vs-U cells; City — TSN "
    "assigns a city code far more aggressively than TSMIS; and Distance To Next Point "
    "— measured to each system's OWN next listed point, and since TSN lists more "
    "breaks its gap is usually smaller (TSN also prints pointer markers \"*P*\" and "
    "\"-------->\" there, conserved verbatim). Read differences in those three as "
    "listing/assignment differences unless the rest of the row agrees.",
)
_write_notes_sheet = ctc.make_notes_writer(_NOTES_TITLE, _NOTES_LINES)


def claims_notes(claims, side_label="TSN"):
    """Human-readable exposure lines for the Notes sheet (CMP-AUD-155): the
    print identity the 12 districts agreed on, the per-route printed
    directions, and the source's own reliability policy."""
    if not claims:
        return [f"{side_label} print: no source-claims record beside this "
                "normalized workbook — rebuild the TSN library to capture the "
                "print identity, per-route directions, and reliability policy."]
    docs = claims.get("documents") or []
    lines = [f"{side_label} print identity: {claims.get('report_id')} "
             f"{claims.get('report_title')} · report {claims.get('report_date')} "
             f"· reference {claims.get('reference_date')} · "
             f"{len(docs)} district print(s)."]
    directions = Counter()
    for d in docs:
        directions.update((d.get("directions") or {}).values())
    if directions:
        summary = " · ".join(f"{k} ×{v}" for k, v in sorted(
            directions.items(), key=lambda kv: (-kv[1], kv[0])))
        lines.append(f"Printed route directions (per district group): {summary}.")
    lines.append(
        "Every district cover carries TSN's own reliability NOTE: landmark "
        "descriptions at Route Breaks/Equates (and possibly county/district "
        "boundaries) may be wrong. Equate rows and blank-county annotations are "
        "therefore disclosed as printed, never repaired.")
    return lines


_SCHEMA = CompareSchema(
    report_name=REPORT_NAME,
    header=SHARED_HEADER,
    side_a="TSMIS",
    side_b="TSN",
    id_noun="location",
    id_noun_plural="locations",
    pair_noun="postmile",
    sides_noun="systems",
    data_widths={"County": 8, "PM": 12, "Description": 26},
    cmp_widths={"PM": 12, "Description": 30},
    one_sided_note_extra=" (mostly TSN segment breaks and TSMIS realignment markers)",
    key_field=KEY_FIELD,
    context_fields=CONTEXT_FIELDS,
    legend_writer=_write_notes_sheet,
    source_file_a=("highway_sequence", TSMIS_SHEET, "xlsx"),   # Source Files sheet
    # HF-09 / PCOA-FINAL-013: say how many of the differing Description cells
    # are the measured representation-only class (the two sources print
    # the same characters once punctuation, spacing, quoting and letter
    # case are set aside). DISCLOSURE ONLY, owner ruling 2026-07-26 -
    # every one of those cells stays flagged and stays inside every
    # published total; the Summary just says how much of the total is
    # presentation rather than data.
    representation_fields=('Description',),
)


def _schema_with_claims(tsn_path, schema=None, title=_NOTES_TITLE,
                        lines=_NOTES_LINES, disclosure=None):
    """The per-run schema: the flavor's static Notes plus the normalized
    workbook's persisted source claims (read from its sidecar — absent claims
    get an explicit rebuild hint instead of silence). CMP-AUD-155. `disclosure`
    is an optional run-resolved callable (the equate-seat counts) that reaches
    both the Summary's disclosure notes and the Notes sheet."""
    base = schema if schema is not None else _SCHEMA
    claim_lines = claims_notes(
        consolidation_meta.read_extra(Path(tsn_path), "tsn_source_claims"))
    notes = tuple(lines) + tuple(claim_lines)
    if disclosure is None:
        return replace(base, legend_writer=ctc.make_notes_writer(title, notes))
    return replace(base, disclosure_notes=(disclosure,),
                   legend_writer=ctc.make_notes_writer(title, notes + (disclosure,)))


# --------------------------------------------------------------------------- #
# adapter surface
# --------------------------------------------------------------------------- #
def suggest_name(tsmis_path):
    return suggest_route_name(tsmis_path, "Highway_Sequence",
                              "TSMIS_vs_TSN_HighwaySequence")


def _load_pair(tsmis_path, tsn_path, stats=None):
    """(rows_t, rows_n, warnings) for the shared driver — no input warnings here, so
    run_compare uses its () default. The Excel export's equate suffix is seated
    the TSN way BEFORE the keys are built (the equate-seat section): TSN
    declares the relations, `stats` receives the seat counts. The evidence
    adapter loads through this same pair, so evidence and comparison can never
    disagree about a row's key."""
    raw_t, _ = _load_tsmis_raw(tsmis_path)
    rows_n, _ = _load_tsn(tsn_path)
    seated = seat_equate_suffixes(raw_t, equate_relations(rows_n), stats)
    return [_tsmis_row(r) for r in seated], rows_n, None


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Highway Sequence TSMIS-vs-TSN comparison workbook(s). `tsmis_path`
    is the consolidated TSMIS Highway Sequence workbook; `tsn_path` the normalized
    TSN workbook (from consolidate_tsn_highway_sequence)."""
    stats = {}          # per-CALL seat counts; resolved when the Summary/Notes are written
    return ctc.run_files_compare(
        _schema_with_claims(tsn_path, disclosure=_equate_seat_disclosure(stats)),
        tsmis_path, tsn_path, out_path,
        banner="Highway Sequence Comparison — TSMIS vs TSN", has_route=True,
        loader=lambda a, b: _load_pair(a, b, stats), deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (openpyxl).",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
