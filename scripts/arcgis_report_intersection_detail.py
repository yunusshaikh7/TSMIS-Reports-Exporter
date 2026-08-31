"""Build OUR Intersection Detail report from the ArcGIS layer library.

The second report rendered this way (Highway Detail was the first, v0.39.0), and
the recipe held: a mapping table plus its own measured rules, not a second
engine. What changes is the SHAPE of the thing being reproduced.

Highway Detail is a SPAN report — the CA HIGHWAYS table printed — so it needs the
span engine, a segmentation, and a merge rule to decide where one record stops.
Intersection Detail is a POINT report: an intersection is a place, not a stretch,
and `IM Intersection Detail` already holds one row per intersection with the
report's identity on it. So there is no segmentation and no merge pass here, and
the projection is nearly direct. Three passes:

  1. the row universe — the IM rows carrying an inventory block;
  2. the approach join, for the ten ML/CS attribute columns and the two lengths;
  3. the county+PM point overlay, for the three columns the intersection row
     does not carry (H/G, City Code, R/U).

Every rule below was measured against the real 2026-08-28 statewide export on
15,154 one-to-one paired intersections, not fitted on a probe case (lessons.md
#17). Two are worth stating outright because a plausible alternative loses:

  * ROW UNIVERSE. The layer holds 38,914 intersection rows, but 22,697 are
    IM-managed shells with the whole inventory half null — no route, no postmile,
    no attributes. The report's universe is the rows with a Main postmile that
    are current (`LRS_DATE_RETIRE` and `InventoryItemEndDate` both null): 16,147,
    against the export's 16,394.

  * WHICH LEG IS PRINTED. The report prints ONE mainline and ONE cross-street
    value for mast arm, channelization, flow, lanes and length, while the layers
    carry ~4.3 approach legs per intersection. `LEG_TYPE` decides it: Major is
    the mainline, Minor is the cross street. Scored across the population, that
    assignment lands 99.4-100.0% on all twelve columns; the reverse assignment
    lands 61.7-95.9%. A rule, not a fit.

`Int St Eff-Date` is `InventoryItemStartDate` on the intersection row — 100.0%,
exact on every paired row. It was very nearly written off as sourceless, which is
the DA2 mistake exactly (`RU Eff` was declared unsourced for Highway Detail and
was sitting in `SHS Population.InventoryItemStartDate` the whole time). Every
column this report prints has a real source; there are no context columns.

Vintage matters here as much as it does for Highway Detail: the build is a
reconstruction AS OF a chosen date, so the as-of must match the export day being
compared or the comparison measures network change instead of correctness.

Console-free; returns a ConsolidateResult. openpyxl loads lazily.
"""
from __future__ import annotations

import logging
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import city_codes
import clean_road_layers as crl
import consolidate_clean_highway as cch
import consolidation_meta
import intersection_detail_columns as idc
import outcome
import paths
from events import ConsolidateResult, Events

log = logging.getLogger(__name__)

REPORT_NAME = "Intersection Detail (ArcGIS)"
SHEET_NAME = "Intersection Detail"      # matches the consolidated export's sheet
MARKER_SHEET = "ArcGIS Report Build"
PROVENANCE_SHEET = "Provenance"
BUILD_VERSION = 1
FILENAME = "intersection_detail_from_layers.xlsx"
OUT_DIR = paths.OUTPUT_ROOT / "arcgis_reports"
OUT_PATH = OUT_DIR / FILENAME

ROUTE_COL = "Route"
HEADER = [ROUTE_COL] + list(idc.HEADER)

# --------------------------------------------------------------------------- #
# The layers this report is built from. All must be staged or the build refuses,
# naming exactly what is missing (the clean-road convention).
# --------------------------------------------------------------------------- #
INX_LAYER = "IM Intersection Detail"
SEG_LAYER = "IM Intersection Approach Segments"
APP_LAYER = "IM Intersection Approach Detail"

# The three columns the intersection row does NOT carry: they belong to the
# highway the intersection sits on, so they come off the same span layers the
# Clean Road build uses, sampled AT the intersection's postmile.
#
# The third element is whether a GAP carries forward. The span layers do not
# tile a route completely, and what the report prints in a gap differs by the
# KIND of attribute — which is a real domain distinction, and measured:
#
#                     strict coverage    carry forward
#     HG                     90.7%           94.5%
#     R/U                    84.0%           91.3%
#     City Code              98.1%           74.9%
#
# Highway Group and Rural/Urban are CONTINUOUS properties of the highway: they
# persist until something changes them, so an uncovered stretch inherits the
# span that ended before it. A city is CONTAINMENT — outside the city limits
# there is no city — so carrying one forward across a gap invents one, which is
# why strict wins there by 23 points. Scored statewide on 15,121 paired
# intersections against the 2026-08-28 export.
OVERLAY_LAYERS = {
    "HG": ("SHS Highway Group", "Highway_Group", True),
    "CITY": ("City", "City_Code", False),
    "POP": ("SHS Population", "Population_Code", True),
}
INTERSECTION_LAYERS = tuple(sorted(
    {INX_LAYER, SEG_LAYER, APP_LAYER}
    | {spec[0] for spec in OVERLAY_LAYERS.values()}))

_SPAN_ID = ["District", "RouteNum", "RouteSuffix", "Alignment", "BeginCounty",
            "EndCounty", "BeginPMPrefix", "EndPMPrefix", "BeginPMMeasure",
            "EndPMMeasure", "LRSFromDate", "LRSToDate", "RouteID"]

_INX_COLS = [
    "INTERSECTION_ID", "LRS_DATE_RETIRE", "InventoryItemStartDate",
    "InventoryItemEndDate", "County_Code", "District_Code",
    "Main_RouteNum", "Main_RouteSuffix", "Main_PMPrefix", "Main_PMSuffix",
    "Main_PMMeasure", "Main_Begin_Date",
    "Cross_RouteNum", "Cross_RouteSuffix", "Cross_PMPrefix", "Cross_PMSuffix",
    "Cross_PMMeasure", "Cross_Begin_Date",
    "Intersection_Name", "Intersection_Geometry", "Int_Geometry_Begin_Date",
    "Intersection_Control", "Int_Control_Begin_Date",
    "Intersection_Lighted_Ind", "Int_Lighted_Ind_Begin_Date",
    "Int_Date_Of_Record",
]
_SEG_COLS = ["INTERSECTION_ID", "APPROACH_ID", "LEG_TYPE"]
_APP_COLS = ["APPROACH_ID", "Number_Thru_Lanes", "FlowCode", "N_Distance",
             "Left_Channel", "Right_Channel_Ind", "Signal_Arm_Ind"]

# The measured leg assignment (see the module docstring).
MAIN_LEG, CROSS_LEG = "Major", "Minor"

# --------------------------------------------------------------------------- #
# The projection: report column -> (layer field(s), how). Written by the
# consolidated export's POSITION, because the export's header LABELS sit shifted
# against their values (the label "INT Type" stands over the eff-date) and the
# comparators read both sides by position. `_POS` is that contract.
# --------------------------------------------------------------------------- #
_POS = {
    "P": 1, "Post Mile": 2, "S": 3, "Location": 4, "Date of Record": 5,
    "H/G": 6, "City Code": 7, "R/U": 8,
    "INT Eff-Date": 9, "INT Type": 10,
    "Ctrl Eff-Date": 11, "Ctrl Type": 12,
    "Light Eff-Date": 13, "Light T/Y": 14,
    "ML Eff-Date": 15, "ML S/M": 16, "ML L/C": 17, "ML R/C": 18,
    "ML T/P": 19, "ML N/L": 20,
    "Description": 21, "Main Line Lgth": 22,
    "Inter Eff-Date": 23, "Inter S": 24, "Inter L": 25, "Inter R": 26,
    "Inter T": 27, "Inter N": 28,
    "Int St Eff-Date": 29,
    "Intrte S": 30, "Intrte Route": 31, "Intrte Post": 32, "Intrte Mile": 33,
    "Xing P/S": 34, "Xing Line Lgth": 35,
}
assert sorted(_POS.values()) == list(range(1, len(HEADER))), \
    "the projection must write every column position but Route"

_MAJOR = f"{APP_LAYER} ({MAIN_LEG} leg)"
_MINOR = f"{APP_LAYER} ({CROSS_LEG} leg)"
PROJECTION = {
    "P": (f"{INX_LAYER}.Main_PMPrefix", "code"),
    "Post Mile": (f"{INX_LAYER}.Main_PMMeasure", "zero-padded 3 decimals"),
    "S": (f"{INX_LAYER}.Main_RouteSuffix", "code"),
    "Location": (f"{INX_LAYER}.District_Code + County_Code + Main_RouteNum",
                 "'DD CCC RRR', county to its TASAS code"),
    "Date of Record": (f"{INX_LAYER}.Int_Date_Of_Record", "date"),
    "H/G": ("SHS Highway Group.Highway_Group", "span sampled at the postmile"),
    "City Code": ("City.City_Code", "span sampled at the postmile, TASAS code"),
    "R/U": ("SHS Population.Population_Code", "span sampled at the postmile"),
    "INT Eff-Date": (f"{INX_LAYER}.Int_Geometry_Begin_Date", "date"),
    "INT Type": (f"{INX_LAYER}.Intersection_Geometry", "code"),
    "Ctrl Eff-Date": (f"{INX_LAYER}.Int_Control_Begin_Date", "date"),
    "Ctrl Type": (f"{INX_LAYER}.Intersection_Control", "code"),
    "Light Eff-Date": (f"{INX_LAYER}.Int_Lighted_Ind_Begin_Date", "date"),
    "Light T/Y": (f"{INX_LAYER}.Intersection_Lighted_Ind", "Yes/No -> Y/N"),
    "ML Eff-Date": (f"{INX_LAYER}.Main_Begin_Date", "date"),
    "ML S/M": (f"{_MAJOR}.Signal_Arm_Ind", "Yes/No -> Y/N"),
    "ML L/C": (f"{_MAJOR}.Left_Channel", "code"),
    "ML R/C": (f"{_MAJOR}.Right_Channel_Ind", "Yes/No -> Y/N"),
    "ML T/P": (f"{_MAJOR}.FlowCode", "code"),
    "ML N/L": (f"{_MAJOR}.Number_Thru_Lanes", "integer"),
    "Description": (f"{INX_LAYER}.Intersection_Name",
                    "text, upper-cased, runs of spaces collapsed"),
    "Main Line Lgth": (f"{_MAJOR}.N_Distance", "integer"),
    "Inter Eff-Date": (f"{INX_LAYER}.Cross_Begin_Date", "date"),
    "Inter S": (f"{_MINOR}.Signal_Arm_Ind", "Yes/No -> Y/N"),
    "Inter L": (f"{_MINOR}.Left_Channel", "code"),
    "Inter R": (f"{_MINOR}.Right_Channel_Ind", "Yes/No -> Y/N"),
    "Inter T": (f"{_MINOR}.FlowCode", "code"),
    "Inter N": (f"{_MINOR}.Number_Thru_Lanes", "integer"),
    "Int St Eff-Date": (f"{INX_LAYER}.InventoryItemStartDate", "date"),
    "Intrte S": (f"{INX_LAYER}.Cross_RouteSuffix", "code"),
    "Intrte Route": (f"{INX_LAYER}.Cross_RouteNum", "integer"),
    "Intrte Post": (f"{INX_LAYER}.Cross_PMPrefix", "code"),
    "Intrte Mile": (f"{INX_LAYER}.Cross_PMMeasure", "number"),
    "Xing P/S": (f"{INX_LAYER}.Cross_PMSuffix", "code"),
    "Xing Line Lgth": (f"{_MINOR}.N_Distance", "integer"),
}
assert set(PROJECTION) == set(_POS), "PROJECTION and _POS must agree"

# This report has no column without a source (see the docstring on Int St
# Eff-Date). Kept as an explicit empty tuple so the comparison, the marker sheet
# and `check_arcgis_report` all read the same contract Highway Detail exposes.
CONTEXT_COLUMNS = ()


# --------------------------------------------------------------------------- #
# cell normalizations (one home; the Provenance sheet names them)
# --------------------------------------------------------------------------- #
def _s(v):
    return "" if v is None else str(v).strip()


def _dec(v):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):  # silent-ok: a pure numeric predicate — a non-numeric cell simply has no number, and every caller renders that as the report's blank
        return None


def _pm3(v):
    """A postmile to the report's fixed zero-padded form ('000.204')."""
    d = _dec(v)
    return "" if d is None else f"{d:07.3f}"


def _date(v):
    """A layer date to the report's 'YY-MM-DD' print form."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%y-%m-%d")
    return _s(v)


def _int(v):
    """A count/length to the report's plain-integer form ('3', '250')."""
    d = _dec(v)
    if d is None:
        return _s(v)
    return str(int(d))


def _yn(v):
    """The layer's Yes/No indicator as the report's Y/N."""
    return {"YES": "Y", "NO": "N", "TRUE": "Y", "FALSE": "N"}.get(
        _s(v).upper(), _s(v))


def _text(v):
    """A description as the report prints it: upper-case, runs of whitespace
    collapsed. The layer pads some names out with runs of spaces the report
    does not render ('PRIVATE RD      LT'); collapsing is what takes this
    column from 98.2% to agreement on the real text."""
    return " ".join(_s(v).upper().split())


def _num(v):
    """A postmile-shaped number as the report prints it (no padding)."""
    d = _dec(v)
    if d is None:
        return _s(v)
    return str(int(d)) if d == int(d) else f"{d.normalize():f}"


# --------------------------------------------------------------------------- #
# the point overlay: what a span layer says AT one postmile
# --------------------------------------------------------------------------- #
class _PointOverlay:
    """The as-of spans of one layer, indexed for `value_at`.

    Keyed the way every clean-road contract is keyed — (route, county, PM
    prefix) — because the same postmile number recurs on a different county or
    a realignment prefix. A cross-county span is indexed under its BEGIN county
    only: its end postmile lives in the next county's space, so its extent there
    is not comparable, and an intersection in that next county is covered by
    that county's own span. Where spans overlap, the newest LRS date wins (the
    span engine's own rank)."""

    __slots__ = ("_by_key", "_carry")

    def __init__(self, carry=False):
        self._by_key = defaultdict(list)
        self._carry = carry

    def add(self, route, county, prefix, begin, end, value, rank):
        self._by_key[(route, county, prefix)].append((begin, end, rank, value))

    def finish(self):
        for spans in self._by_key.values():
            spans.sort(key=lambda t: (t[0], t[1]))

    def value_at(self, route, county, prefix, pm):
        """The covering span's value, or "" where nothing covers the point.

        Half-open on the end so a point on a boundary belongs to the span that
        STARTS there, matching how the report's own records begin. Where spans
        overlap the newest LRS date wins.

        When this layer carries (see OVERLAY_LAYERS), an uncovered point falls
        back to the span that ended nearest before it — the attribute persists
        until something changes it. A non-carrying layer answers "" instead,
        because for containment attributes a gap means genuinely nothing."""
        best, best_rank = "", None
        carried, carried_end = "", None
        for begin, end, rank, value in self._by_key.get(
                (route, county, prefix), ()):
            if begin > pm:
                break
            if begin <= pm < end or (begin == pm == end):
                if best_rank is None or rank > best_rank:
                    best, best_rank = value, rank
            elif self._carry and (carried_end is None or end >= carried_end):
                carried, carried_end = value, end
        if best or not self._carry:
            return best
        return carried


def _read_overlay(lib, index, tag, *, asof, events):
    """One span layer's as-of spans as a `_PointOverlay`."""
    layer, attr, carry = OVERLAY_LAYERS[tag]
    entry = index.get(layer) or {}
    ov = _PointOverlay(carry)
    n = 0
    for r in crl.stream_layer(lib["present"][layer], _SPAN_ID + [attr],
                              layer_name=layer,
                              expected_rows=entry.get("rows"),
                              optional=("InventoryItemStartDate",)):
        rid = r["RouteID"]
        if isinstance(rid, str) and rid and not rid.startswith("SHS_"):
            continue                       # all-roads layers: SHS routes only
        if not crl.is_asof(crl.to_serial(r["LRSFromDate"]),
                           crl.to_serial(r["LRSToDate"]), asof):
            continue
        b, e = crl.pm_units(r["BeginPMMeasure"]), crl.pm_units(r["EndPMMeasure"])
        if b is None or e is None:
            continue                       # unplaceable without a postmile
        route = _route_token(r["RouteNum"], r["RouteSuffix"])
        county = crl.norm_county(r["BeginCounty"])
        prefix = crl.dot_none(r["BeginPMPrefix"]).strip().upper()
        if crl.norm_county(r["EndCounty"]) != county or e < b:
            e = b                          # cross-county: its extent here is
        n += 1                             # only the point it begins at
        ov.add(route, county, prefix, b, e,
               _s(r[attr]), crl.to_serial(r["LRSFromDate"]) or 0.0)
    ov.finish()
    events.on_log(f"  {layer}: {n:,} as-of spans")
    return ov


def _route_token(route_num, route_suffix):
    """The consolidated export's route identity ('001', '008U').

    The suffix goes through `dot_none`: the layer exports write '.' for "no
    suffix", so taking it verbatim builds '001.' and matches nothing."""
    return crl.norm_route(route_num) + crl.dot_none(route_suffix).strip().upper()


# --------------------------------------------------------------------------- #
# the build
# --------------------------------------------------------------------------- #
def _read_legs(lib, index, events):
    """{intersection id: {LEG_TYPE: [approach attribute dict]}} — the join that
    supplies the ML/CS block."""
    app = {}
    entry = index.get(APP_LAYER) or {}
    for r in crl.stream_layer(lib["present"][APP_LAYER], _APP_COLS,
                              layer_name=APP_LAYER,
                              expected_rows=entry.get("rows")):
        app[r["APPROACH_ID"]] = r
    entry = index.get(SEG_LAYER) or {}
    legs = defaultdict(lambda: defaultdict(list))
    joined = 0
    for r in crl.stream_layer(lib["present"][SEG_LAYER], _SEG_COLS,
                              layer_name=SEG_LAYER,
                              expected_rows=entry.get("rows")):
        detail = app.get(r["APPROACH_ID"])
        if detail is None:
            continue
        joined += 1
        legs[r["INTERSECTION_ID"]][_s(r["LEG_TYPE"])].append(detail)
    events.on_log(f"  {APP_LAYER}: {len(app):,} approaches, "
                  f"{joined:,} joined to an intersection")
    return legs


def _leg_value(legs_for_type, field, conv):
    """The printed value of one approach attribute.

    The legs of one type usually agree (both mainline directions carry the same
    inventory), so the common case is unambiguous. Where they disagree the most
    frequent value is printed, ties broken by the value's own order, so the
    build is deterministic rather than dependent on layer row order."""
    if not legs_for_type:
        return ""
    counts = defaultdict(int)
    for leg in legs_for_type:
        counts[conv(leg[field])] += 1
    return max(sorted(counts), key=lambda v: (counts[v], v != "", v))


def _project(r, legs, overlays):
    """The printed value of every report column for one intersection row."""
    route = _route_token(r["Main_RouteNum"], r["Main_RouteSuffix"])
    county = crl.norm_county(r["County_Code"])
    prefix = crl.dot_none(r["Main_PMPrefix"]).strip().upper()
    pm_units = crl.pm_units(r["Main_PMMeasure"])
    district = crl.norm_district(r["District_Code"])
    major = legs.get(MAIN_LEG, ())
    minor = legs.get(CROSS_LEG, ())

    def at(tag):
        if pm_units is None:
            return ""
        return overlays[tag].value_at(route, county, prefix, pm_units)

    return {
        "P": prefix,
        "Post Mile": _pm3(r["Main_PMMeasure"]),
        "S": crl.dot_none(r["Main_RouteSuffix"]).strip().upper(),
        "Location": f"{district} {county} {crl.norm_route(r['Main_RouteNum'])}",
        "Date of Record": _date(r["Int_Date_Of_Record"]),
        "H/G": crl.code_of(at("HG")),
        "City Code": city_codes.norm_city(at("CITY")),
        "R/U": crl.code_of(at("POP")),
        "INT Eff-Date": _date(r["Int_Geometry_Begin_Date"]),
        "INT Type": crl.code_of(r["Intersection_Geometry"]),
        "Ctrl Eff-Date": _date(r["Int_Control_Begin_Date"]),
        "Ctrl Type": crl.code_of(r["Intersection_Control"]),
        "Light Eff-Date": _date(r["Int_Lighted_Ind_Begin_Date"]),
        "Light T/Y": _yn(r["Intersection_Lighted_Ind"]),
        "ML Eff-Date": _date(r["Main_Begin_Date"]),
        "ML S/M": _leg_value(major, "Signal_Arm_Ind", _yn),
        "ML L/C": _leg_value(major, "Left_Channel", crl.code_of),
        "ML R/C": _leg_value(major, "Right_Channel_Ind", _yn),
        "ML T/P": _leg_value(major, "FlowCode", crl.code_of),
        "ML N/L": _leg_value(major, "Number_Thru_Lanes", _int),
        "Description": _text(r["Intersection_Name"]),
        "Main Line Lgth": _leg_value(major, "N_Distance", _int),
        "Inter Eff-Date": _date(r["Cross_Begin_Date"]),
        "Inter S": _leg_value(minor, "Signal_Arm_Ind", _yn),
        "Inter L": _leg_value(minor, "Left_Channel", crl.code_of),
        "Inter R": _leg_value(minor, "Right_Channel_Ind", _yn),
        "Inter T": _leg_value(minor, "FlowCode", crl.code_of),
        "Inter N": _leg_value(minor, "Number_Thru_Lanes", _int),
        "Int St Eff-Date": _date(r["InventoryItemStartDate"]),
        "Intrte S": crl.dot_none(r["Cross_RouteSuffix"]).strip().upper(),
        "Intrte Route": _int(r["Cross_RouteNum"]) if _s(r["Cross_RouteNum"]) else "",
        "Intrte Post": crl.dot_none(r["Cross_PMPrefix"]).strip().upper(),
        "Intrte Mile": _num(r["Cross_PMMeasure"]) if _s(r["Cross_PMMeasure"]) else "",
        "Xing P/S": crl.dot_none(r["Cross_PMSuffix"]).strip().upper(),
        "Xing Line Lgth": _leg_value(minor, "N_Distance", _int),
    }


def build_rows(lib, index, *, asof, events):
    """(rows, stats) — the report-shaped rows, one per current inventory-bearing
    intersection."""
    events.on_log("  reading the span layers the report samples...")
    overlays = {tag: _read_overlay(lib, index, tag, asof=asof, events=events)
                for tag in OVERLAY_LAYERS}
    legs = _read_legs(lib, index, events)

    entry = index.get(INX_LAYER) or {}
    rows, seen, shells, retired = [], 0, 0, 0
    for r in crl.stream_layer(lib["present"][INX_LAYER], _INX_COLS,
                              layer_name=INX_LAYER,
                              expected_rows=entry.get("rows")):
        if events.is_cancelled():
            return [], {}
        if _s(r["Main_PMMeasure"]) == "":
            shells += 1           # an IM-managed shell: no inventory block
            continue
        if _s(r["LRS_DATE_RETIRE"]) or _s(r["InventoryItemEndDate"]):
            retired += 1
            continue
        seen += 1
        proj = _project(r, legs.get(r["INTERSECTION_ID"], {}), overlays)
        row = [""] * len(HEADER)
        row[0] = _route_token(r["Main_RouteNum"], r["Main_RouteSuffix"])
        for name, pos in _POS.items():
            row[pos] = proj[name]
        rows.append(row)

    rows.sort(key=lambda r: (r[0], r[_POS["Location"]], r[_POS["Post Mile"]]))
    stats = {"source_rows": seen, "rows": len(rows), "shells": shells,
             "retired": retired, "routes": len({r[0] for r in rows})}
    events.on_log(f"  built {len(rows):,} intersection records "
                  f"({stats['routes']} routes); skipped {shells:,} rows with no "
                  f"inventory block and {retired:,} retired")
    return rows, stats


# --------------------------------------------------------------------------- #
# reading the built workbook
# --------------------------------------------------------------------------- #
def is_arcgis_report(path):
    """Is `path` a workbook THIS module built? Identity is the marker sheet it
    writes — the comparison uses it to enforce which side is which, so a
    missing/unreadable workbook answers False rather than raising."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True)
    except Exception as e:    # silent-ok: a role probe — an unopenable file is simply not one of ours, and the loader that follows reports the real problem
        log.info("arcgis report probe: %s unreadable (%s: %s)",
                 path, type(e).__name__, e)
        return False
    try:
        return MARKER_SHEET in wb.sheetnames
    finally:
        wb.close()


def report_facts(path):
    """The build facts this module's marker sheet records. `{}` when absent or
    unreadable — never invented."""
    from openpyxl import load_workbook

    facts = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if MARKER_SHEET not in wb.sheetnames:
            return facts
        keys = {"As-of date": "asof", "Build version": "build_version",
                "Report records written": "rows", "Routes": "routes",
                "Layer library": "layer_library",
                "Rows with no inventory block": "shells",
                "Retired rows skipped": "retired"}
        for r in wb[MARKER_SHEET].iter_rows(values_only=True):
            if not r or r[0] is None:
                continue
            key = keys.get(_s(r[0]))
            if key:
                facts[key] = _s(r[1] if len(r) > 1 else None)
        return facts
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# workbook
# --------------------------------------------------------------------------- #
def _provenance_rows():
    """(printed label, what it holds, layer field(s), how, tier) — every column
    indexed back to what produced it.

    The label and the content are separate columns on purpose: the export's
    header labels sit shifted against their values (the column LABELLED
    'INT Type' holds the geometry EFFECTIVE DATE), and this build reproduces
    that layout exactly so the comparators can read both sides by position.
    Naming only the label would misdescribe half the sheet."""
    by_pos = {pos: name for name, pos in _POS.items()}
    out = []
    for pos in range(1, len(HEADER)):
        holds = by_pos[pos]
        src, how = PROJECTION[holds]
        tier = "sampled" if holds in ("H/G", "City Code", "R/U") else "projected"
        out.append((HEADER[pos], holds, src, how, tier))
    return out


def _write_workbook(out_path, rows, stats, asof, lib_root):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    head_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", start_color="1F3864")
    ws.append(HEADER)
    for c in next(ws.iter_rows(min_row=1, max_row=1)):
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    widths = {"Route": 7, "Post Mile": 11, "Location": 13, "Description": 26,
              "Date of Record": 11}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 8)

    prov = wb.create_sheet(PROVENANCE_SHEET)
    prov.sheet_properties.tabColor = "1F3864"
    prov.append(["Printed label", "What the column holds",
                 "ArcGIS layer field(s)", "How", "Tier"])
    for c in next(prov.iter_rows(min_row=1, max_row=1)):
        c.font, c.fill = head_font, head_fill
    for row in _provenance_rows():
        prov.append(list(row))
    for col, w in (("A", 16), ("B", 18), ("C", 56), ("D", 40), ("E", 12)):
        prov.column_dimensions[col].width = w

    mk = wb.create_sheet(MARKER_SHEET)
    mk.sheet_properties.tabColor = "1F3864"
    for k, v in [
        ("Build version", BUILD_VERSION),
        ("As-of date", asof),
        ("Layer library", str(lib_root)),
        ("Report records written", stats["rows"]),
        ("Routes", stats["routes"]),
        ("Rows with no inventory block", stats["shells"]),
        ("Retired rows skipped", stats["retired"]),
        ("Row universe", "IM Intersection Detail rows carrying a Main postmile "
                         "whose LRS_DATE_RETIRE and InventoryItemEndDate are "
                         "both empty"),
        ("Leg rule", f"the printed mainline block comes from the "
                     f"{MAIN_LEG} approach legs and the cross-street block from "
                     f"the {CROSS_LEG} legs; where legs of one type disagree the "
                     f"most frequent value is printed"),
        ("Columns with no source", ", ".join(CONTEXT_COLUMNS) or "(none)"),
    ]:
        mk.append([k, v])
    mk.column_dimensions["A"].width = 30
    mk.column_dimensions["B"].width = 70
    for row in mk.iter_rows(min_col=2, max_col=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# public entry point
# --------------------------------------------------------------------------- #
def consolidate(events=None, confirm_overwrite=None, day=None, *,
                out_path=None, asof=None, lib_root=None):
    """Build the Intersection Detail report FROM THE ARCGIS LAYERS.

    `asof` is the reconstruction date — set it to the export day being compared
    against, or the comparison measures network change instead of correctness.
    `out_path` defaults to `output/arcgis_reports/`; the destination is app-owned
    and a rebuild replaces it by design."""
    del day                       # ConsolidateWorker passes it; no run days here
    events = events or Events()
    out_path = Path(out_path) if out_path else OUT_PATH
    try:
        return _consolidate(events, confirm_overwrite, out_path, asof, lib_root)
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))


def _consolidate(events, confirm_overwrite, out_path, asof, lib_root):
    if out_path.exists() and confirm_overwrite is not None \
            and not confirm_overwrite(out_path):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.",
                                 completion=outcome.CANCELLED)

    lib = crl.inventory(lib_root)
    missing = [name for name in INTERSECTION_LAYERS
               if name not in lib["present"]]
    if missing:
        raise ValueError(
            "The ArcGIS layer library is missing the layer(s) this report is "
            "built from:\n\n  " + "\n  ".join(missing)
            + "\n\nDrop those layer exports into the arcgis_layers folder and "
              "build again.")

    index = crl.read_index(lib_root)
    # One date resolver for both builds: a second copy of "None means the
    # staged extract's own date, otherwise parse it" is exactly the kind of
    # duplicate that drifts.
    asof_date = cch._resolve_asof(asof)
    asof_text = asof_date.isoformat()

    events.on_log("=" * 60)
    events.on_log(f"{REPORT_NAME} — built from the ArcGIS layers "
                  f"(as of {asof_text})")
    events.on_log("=" * 60)

    rows, stats = build_rows(lib, index, asof=crl.to_serial(asof_date),
                             events=events)
    if events.is_cancelled():
        return ConsolidateResult(status="cancelled", message="Cancelled.",
                                 completion=outcome.CANCELLED)
    if not rows:
        return ConsolidateResult(
            status="error",
            message="The intersection layers produced no records to project — "
                    "check the layer library and the as-of date.")

    _write_workbook(out_path, rows, stats, asof_text, crl.root())
    result = ConsolidateResult(
        status="ok",
        message=(f"Built {stats['rows']:,} Intersection Detail records "
                 f"({stats['routes']} routes) from the ArcGIS layers as of "
                 f"{asof_text}."),
        output_path=str(out_path),
        summary_lines=[f"{REPORT_NAME}: {stats['rows']:,} records, "
                       f"{stats['routes']} routes -> {out_path.name}"],
        completion=outcome.COMPLETE)
    if not consolidation_meta.write_outcome(
            out_path, result,
            extra={"arcgis_report_build": {
                "report": "intersection_detail",
                "build_version": BUILD_VERSION,
                "asof": asof_text,
                "layer_library": str(crl.root()),
                "context_columns": list(CONTEXT_COLUMNS),
                **stats,
            }}):
        return ConsolidateResult(
            status="error",
            message="The build finished but its outcome sidecar could not be "
                    "published — rebuild before comparing.")
    result.sidecar_published = True
    events.on_log(result.message)
    return result
