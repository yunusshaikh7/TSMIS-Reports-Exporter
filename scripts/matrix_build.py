"""The matrix BUILD side (S4 / ARC-03, split from matrix.py).

(Re)building comparison cells: the live-formulas twin policy, the store-folder
consolidation for the TSN modes, and `build_comparison` — the mode dispatcher
the GUI workers and the one-click validation drive. Collaborators that the
verification suite (or a future caller) monkeypatches on the `matrix` facade
are looked up through the facade AT CALL TIME (`_m.<name>`), so patching
`matrix.read_counts` / `matrix._ensure_consolidated` / … keeps intercepting
these internal calls exactly as it did before the split.
"""
import contextlib
import hashlib
import logging
import os
import shutil
import stat
import tempfile
from pathlib import Path

import consolidation_meta
import outcome
import owned_dir
import reports

import artifact_store
from events import ConsolidateResult
from matrix_state import (_MTIME_TOL_S, _cell_input_fingerprint, _mode_by_id,
                          _pdf_self_comparator, _row_defs, _row_modes,
                          _safe_mtime, comparison_path, mode_out_path,
                          producer_identity, record_result, record_tsn_result,
                          tsn_input_root, tsn_source)

class _FacadeProxy:
    """Resolves `_m.<name>` through the `matrix` facade AT CALL TIME (a lazy
    import, so there is no module-level matrix <-> matrix_build cycle). This is
    what keeps `matrix.<name>` the one true monkeypatch point for the names the
    verification suite stubs."""
    def __getattr__(self, name):
        import matrix
        return getattr(matrix, name)


_m = _FacadeProxy()       # the facade; resolved at call time (see the docstring)

log = logging.getLogger("tsmis.matrix")


def _guard_allows(commit_guard, path=None):
    """Evaluate the Matrix's target-aware ownership callback fail closed."""
    if commit_guard is None:
        return True
    try:
        if path is None:
            return bool(commit_guard())
        return consolidation_meta.guard_allows(commit_guard, Path(path))
    except Exception as e:                       # noqa: BLE001 - fail closed on guard defects
        log.error("matrix output guard raised (%s: %s)", type(e).__name__, e)
        return False


def _require_commit_guard(commit_guard, action, path=None):
    """Fail before a Matrix mutation outside its exact leased destination."""
    if not _guard_allows(commit_guard, path):
        raise ValueError(
            "The Matrix output destination changed while this cell was building. "
            f"It was not used for the {action}; retry the comparison.")


def _compose_source_guard(commit_guard, source_identity_check):
    """Preserve the target lease while also binding exact TSN source bytes."""
    if source_identity_check is None:
        return commit_guard

    def _guard(path=None, **binding):
        if commit_guard is not None:
            try:
                if path is None:
                    target_current = bool(commit_guard())
                elif binding:
                    # Target-aware identity bindings are security constraints,
                    # not optional hints. A legacy/path-only guard cannot certify
                    # an anchor/directory binding by silently dropping kwargs.
                    target_current = bool(commit_guard(Path(path), **binding))
                else:
                    # Backward-compatible path-only guards remain valid when no
                    # target identity binding was supplied.
                    target_current = bool(commit_guard(Path(path)))
            except Exception as e:                   # noqa: BLE001 - fail closed
                log.error("matrix target guard raised (%s: %s)",
                          type(e).__name__, e)
                return False
            if not target_current:
                return False
        try:
            return bool(source_identity_check())
        except Exception as e:                       # noqa: BLE001 - fail closed
            log.error("matrix TSN source guard raised (%s: %s)",
                      type(e).__name__, e)
            return False

    return _guard


def _require_source_identity(source_identity_check, action):
    if source_identity_check is None:
        raise ValueError(
            "The TSN source has no certifiable content identity. Refresh the "
            f"comparison before {action}.")
    try:
        current = bool(source_identity_check())
    except Exception as e:                           # noqa: BLE001 - fail closed
        log.error("matrix TSN identity check raised (%s: %s)",
                  type(e).__name__, e)
        current = False
    if not current:
        raise ValueError(
            "The TSN source generation changed or is no longer certifiably "
            f"current. Refresh the comparison before {action}.")


def tsn_identity_check_for(tsn_key, source):
    """Return ``(token, current_predicate)`` for one resolved TSN source."""
    import tsn_library
    token = source.get("identity_token")
    selection = source.get("selection")
    if source.get("kind") == "file" and selection is not None:
        if token is None:
            raise ValueError(
                "The selected TSN workbook has no certifiable content identity; "
                "re-pick it and refresh the comparison.")

        def _explicit_current():
            try:
                tsn_library.require_explicit_selection(selection)
                return True
            except (OSError, ValueError) as e:
                # The caller turns False into a source-identity refusal, but the
                # WHY only survives in the log ("one log upload answers it").
                log.warning("matrix: selected TSN workbook is no longer current "
                            "(%s: %s)", type(e).__name__, e)
                return False

        _require_source_identity(_explicit_current, "using the selected TSN workbook")
        return token, _explicit_current

    if source.get("kind") == "consolidated":
        if not isinstance(token, str) or not token:
            raise ValueError(
                "The consolidated TSN workbook is stale, legacy, foreign, or has "
                "no canonical content certificate. Refresh the comparison after "
                "rebuilding the authoritative TSN source.")

        def _canonical_current():
            status = tsn_library.status(tsn_key)
            return bool(status.get("current")
                        and status.get("identity_token") == token)

        _require_source_identity(_canonical_current, "using the TSN workbook")
        return token, _canonical_current

    raise ValueError("no certifiably current TSN workbook available")


def require_cached_tsn_identity(record, token):
    """Return the bound comparison generation or reject a mismatched cache."""
    if not isinstance(record, dict):
        raise ValueError(
            "The comparison has no trusted TSN cache record; refresh the "
            "comparison before generating evidence.")
    if record.get("source_identities") != {"tsn": token}:
        raise ValueError(
            "The comparison was built from a different or unrecorded TSN source "
            "generation; refresh the comparison before generating evidence.")
    generation_id = record.get("generation_id")
    if not isinstance(generation_id, str) or not generation_id:
        raise ValueError(
            "The comparison cache is not bound to an artifact generation; refresh "
            "the comparison before generating evidence.")
    return generation_id


def tsn_expected_workbook_identity(tsn_key, source, token=None):
    """Return the exact content identity the attempt-local copy must match."""
    import tsn_library
    token = source.get("identity_token") if token is None else token
    if source.get("kind") == "file" and source.get("selection") is not None:
        if not isinstance(token, dict):
            raise ValueError("the selected TSN workbook identity is unavailable")
        try:
            expected = {
                "version": 1,
                "algorithm": "sha256",
                "byte_length": token["size"],
                "sha256": token["sha256"],
            }
            return tsn_library.validate_normalized_workbook_identity(expected)
        except (KeyError, TypeError, ValueError) as e:
            raise ValueError(
                "the selected TSN workbook content identity is invalid; re-pick it") from e
    if source.get("kind") == "consolidated":
        status = tsn_library.status(tsn_key)
        if (not status.get("current")
                or status.get("identity_token") != token):
            raise ValueError(
                "the canonical TSN workbook changed before its content could be captured")
        try:
            return tsn_library.validate_normalized_workbook_identity(
                status.get("normalized_workbook_identity"))
        except ValueError as e:
            raise ValueError(
                "the canonical TSN workbook has no valid content identity") from e
    raise ValueError("no certifiably current TSN workbook is available to capture")


def _stat_signature(st):
    return (int(st.st_size),
            int(getattr(st, "st_mtime_ns", st.st_mtime * 1e9)),
            int(getattr(st, "st_dev", 0)), int(getattr(st, "st_ino", 0)))


def _object_identity(st):
    return (int(getattr(st, "st_dev", 0)), int(getattr(st, "st_ino", 0)),
            stat.S_IFMT(st.st_mode))


@contextlib.contextmanager
def captured_tsn_workbook(source_path, expected_identity):
    """Yield a private immutable-by-ownership copy bound to expected source bytes.

    The source descriptor and its pathname must identify the same stable file
    before and after copying. The private copy is then independently re-hashed
    and must equal the content identity captured from resolution/certification.
    """
    import tsn_library
    expected = tsn_library.validate_normalized_workbook_identity(expected_identity)
    source = Path(source_path)
    temp_root = Path(tempfile.mkdtemp(prefix="tsmis-tsn-consumer-"))
    try:
        temp_identity = _object_identity(temp_root.lstat())
    except OSError as e:
        raise ValueError("the private TSN capture directory could not be bound") from e
    captured = temp_root / source.name
    captured_identity = None
    captured_sidecar = consolidation_meta.meta_path(captured)
    sidecar_identity = None
    try:
        source_outcome_before = consolidation_meta.read_outcome(source)
        descriptor = None
        try:
            bound_path, descriptor, opened = tsn_library._open_bound_file(source)
            digest = hashlib.sha256()
            count = 0
            with captured.open("xb") as writer:
                captured_identity = _object_identity(os.fstat(writer.fileno()))
                while True:
                    block = os.read(descriptor, 1024 * 1024)
                    if not block:
                        break
                    writer.write(block)
                    digest.update(block)
                    count += len(block)
            tsn_library._finish_bound_file(bound_path, descriptor, opened)
            descriptor = None
        except (OSError, ValueError) as e:
            if descriptor is not None:
                try:
                    os.close(descriptor)
                except OSError:  # silent-ok: cleanup close; the ValueError below carries the real cause
                    pass
            raise ValueError(
                "the TSN workbook could not be captured through a stable pathname "
                f"({type(e).__name__}: {e})") from e
        copied_identity = {
            "version": 1, "algorithm": "sha256",
            "byte_length": count, "sha256": digest.hexdigest(),
        }
        if copied_identity != expected:
            raise ValueError(
                "the TSN workbook bytes did not match the resolved source generation")
        # Re-read the private pathname too: a local alias/replacement between copy
        # close and comparator entry cannot inherit the source claim.
        if tsn_library.normalized_workbook_identity(captured) != expected:
            raise ValueError("the private TSN workbook capture could not be verified")
        source_outcome_after = consolidation_meta.read_outcome(source)

        def _outcome_contract(record):
            if record is None:
                return None
            return (record.completion, record.skipped_inputs,
                    record.failed_inputs, record.trusted, record.current)

        if (_outcome_contract(source_outcome_before)
                != _outcome_contract(source_outcome_after)):
            raise ValueError(
                "the TSN workbook producer outcome changed during capture")
        if source_outcome_after is not None:
            if not source_outcome_after.trusted or not source_outcome_after.current:
                raise ValueError(
                    "the TSN workbook producer outcome is not trusted")
            snapshot_result = ConsolidateResult(
                status="ok", output_path=str(captured),
                completion=source_outcome_after.completion,
                skipped_inputs=source_outcome_after.skipped_inputs or 0,
                failed_inputs=source_outcome_after.failed_inputs or 0)
            if not consolidation_meta.write_outcome(captured, snapshot_result):
                raise ValueError(
                    "the TSN workbook producer outcome could not be bound to its capture")
            try:
                sidecar_identity = _object_identity(captured_sidecar.lstat())
            except OSError as e:
                raise ValueError(
                    "the captured TSN producer outcome could not be verified") from e
        yield captured
    finally:
        # Never recursively clean an untrusted/replaced directory. Remove only
        # the exact file and directory objects created by this attempt.
        if sidecar_identity is not None:
            try:
                if _object_identity(captured_sidecar.lstat()) == sidecar_identity:
                    captured_sidecar.unlink()
                else:
                    log.warning("matrix: retained replaced TSN capture sidecar %s",
                                captured_sidecar)
            except FileNotFoundError:  # silent-ok: already gone is the desired end state
                pass
            except OSError as e:
                log.warning("matrix: could not clean TSN capture sidecar %s (%s: %s)",
                            captured_sidecar, type(e).__name__, e)
        if captured_identity is not None:
            try:
                if _object_identity(captured.lstat()) == captured_identity:
                    captured.unlink()
                else:
                    log.warning("matrix: retained replaced TSN capture %s", captured)
            except FileNotFoundError:  # silent-ok: already gone is the desired end state
                pass
            except OSError as e:
                log.warning("matrix: could not clean TSN capture %s (%s: %s)",
                            captured, type(e).__name__, e)
        try:
            if _object_identity(temp_root.lstat()) == temp_identity:
                os.rmdir(temp_root)
            else:
                log.warning("matrix: retained replaced TSN capture directory %s",
                            temp_root)
        except FileNotFoundError:  # silent-ok: already gone is the desired end state
            pass
        except OSError as e:
            log.warning("matrix: retained non-empty/uncertain TSN capture directory "
                        "%s (%s: %s)", temp_root, type(e).__name__, e)


def _partial_comparison_reason(typed):
    """Human reason for a typed non-certifying comparison outcome."""
    if getattr(typed, "pairing_quality", None) == "capped":
        return (
            "duplicate-row identity exceeded the exact-pairing limit; re-scope "
            "and refresh the comparison before generating current evidence")
    if (getattr(typed, "coverage_diagnostics", ())
            or getattr(typed, "warnings", ())
            or getattr(typed, "failures", ())):
        return (
            "the comparison used incomplete inputs — refresh it before "
            "generating current evidence")
    return (
        "the comparison outcome is partial/non-certifying — refresh it before "
        "generating current evidence")


# --------------------------------------------------------------------------- #
# optional live-formulas twin (opt-in). The matrix's offline counts + freshness
# all key off the VALUES workbook at the canonical out_path (compare_core is
# regression-locked, so we can't make mode="both" put values there). So when the
# user opts in, we ALSO write a recalculating formulas copy to a "(formulas)"
# sibling via a second compare pass — best-effort, never failing the values cell.
# --------------------------------------------------------------------------- #
def _formulas_sibling(out_path):
    out_path = Path(out_path)
    return out_path.with_name(f"{out_path.stem} (formulas){out_path.suffix}")


# The live-formulas twin rewrites the whole comparison as live formulas; for the largest
# report (Intersection Detail, ~17k rows) that is millions of formulas and minutes of
# wall-clock ON TOP OF the values workbook that already holds every value. Past this many
# Comparison-sheet rows the matrix skips the twin and says so — keeping the bulk rebuild
# responsive. (The manual Compare tab path doesn't go through here, so an explicitly
# requested single live-formulas comparison is unaffected.)
_FORMULAS_TWIN_MAX_ROWS = 12_000


def _comparison_row_count(values_path):
    """Data-row count of a values workbook's Comparison sheet (read-only — the dimension
    is read from the sheet, not by scanning cells), or None if it can't be read. None ⇒
    the caller writes the twin anyway (never skip the twin on an uncertain probe)."""
    from openpyxl import load_workbook           # lazy: openpyxl stays off GUI startup
    try:
        wb = load_workbook(values_path, read_only=True)
    except Exception as e:                       # noqa: BLE001 (best-effort probe)
        log.debug("formulas-twin probe: can't open %s (%s: %s)",
                  values_path, type(e).__name__, e)
        return None
    try:
        ws = wb["Comparison"] if "Comparison" in wb.sheetnames else wb.worksheets[0]
        n = ws.max_row
    except Exception as e:                       # noqa: BLE001
        log.debug("formulas-twin probe: can't size %s (%s: %s)",
                  values_path, type(e).__name__, e)
        n = None
    finally:
        try:
            wb.close()
        except Exception:                        # noqa: BLE001
            pass
    return (n - 1) if isinstance(n, int) and n > 0 else None   # minus the header row


def _try_formulas(compare_call, out_path, events=None, source_paths=(),
                  commit_guard=None):
    """Run `compare_call(formulas_path)` (mode='formulas') beside the values copy,
    through the adapter's own atomic transaction, so an interrupted formulas write
    never truncates a prior good sibling
    (F9). Best-effort: a failure here must NOT fail the already-written values cell — but it
    is LOGGED (P2-A03: a validation/finalization failure RETURNS status='error' rather than
    raising, so the returned result is inspected, not just exceptions).

    Returns True iff a FRESH twin was committed (CMP-AUD-082 — the caller clears a
    stale prior twin when this is False). Skipped for very large comparisons
    (`_FORMULAS_TWIN_MAX_ROWS`): there the twin is millions of formulas and minutes of
    work, and the values workbook — already committed — holds every value. The skip is
    announced (events + log) so the absent twin is never a silent surprise."""
    rows = _comparison_row_count(out_path)
    # read the limit through the facade — it's a TUNABLE the twin-guard check
    # (and a hotfix) can set as matrix._FORMULAS_TWIN_MAX_ROWS
    if rows is not None and rows > _m._FORMULAS_TWIN_MAX_ROWS:
        msg = (f"Skipping the live-formulas copy of {Path(out_path).name} "
               f"({rows:,} rows, over the {_m._FORMULAS_TWIN_MAX_ROWS:,}-row limit). The "
               f"values workbook has every value; build a live-formulas copy of this "
               f"comparison on its own if you need one.")
        if events is not None:
            events.on_log(msg)
        log.info("matrix: %s", msg)
        return False
    try:
        formulas_path = _formulas_sibling(out_path)
        _require_commit_guard(commit_guard, "live-formulas write", formulas_path)
        artifact_store.ensure_outputs_do_not_alias_sources(
            (formulas_path,), source_paths)
        res = compare_call(formulas_path)
        if getattr(res, "status", None) != "ok":
            log.warning("matrix: live-formulas workbook for %s not refreshed (%s)",
                        Path(out_path).name, getattr(res, "message", "") or "commit failed")
            return False
        return True
    except Exception as e:                       # noqa: BLE001
        log.warning("matrix: live-formulas workbook for %s not written (%s: %s)",
                    Path(out_path).name, type(e).__name__, e)
        return False


def _clear_stale_formulas_twin(out_path, events=None, commit_guard=None,
                               source_paths=()):
    """Remove a `(formulas)` sibling that is NOT being refreshed beside a just-committed
    values workbook (CMP-AUD-082).

    The values workbook is canonical and the optional live-formulas twin has no manifest
    or freshness state, so a values-only refresh, an inputs-changed skip, an over-limit
    skip, or a failed formulas commit would otherwise leave an ordinary audit-looking
    `(formulas)` file from an OLDER generation looking current beside the newer values
    comparison. Removing it is the safe resolution (the values workbook holds every
    value; a genuine live-formulas copy is one explicit rebuild away). Best-effort +
    ownership/alias-guarded: a twin still open in Excel (locked) can't be removed, so it
    is announced instead of silently trusted. No-op when no prior twin exists (the common
    first-build case)."""
    formulas_path = _formulas_sibling(out_path)
    try:
        if not formulas_path.exists():
            return
    except OSError:  # silent-ok: an unstattable path is treated as absent
        return
    if not _guard_allows(commit_guard, formulas_path):
        log.warning("matrix: a stale live-formulas twin %s was left (destination "
                    "ownership changed); it does not match the current values workbook",
                    formulas_path.name)
        return
    try:
        # Never delete a selected comparison source that shares the twin's name.
        artifact_store.ensure_outputs_do_not_alias_sources(
            (formulas_path,), source_paths)
    except ValueError:
        log.warning("matrix: retained a live-formulas twin %s that aliases a source",
                    formulas_path.name)
        return
    try:
        formulas_path.unlink()
    except FileNotFoundError:  # silent-ok: concurrent absence is the goal
        return
    except OSError as e:  # locked (open in Excel) or unremovable — announce, keep going
        log.warning("matrix: could not remove the stale live-formulas twin %s (%s: %s; "
                    "open in Excel?); it does not match the current values workbook",
                    formulas_path.name, type(e).__name__, e)
        return
    msg = (f"Removed the prior live-formulas copy of {Path(out_path).name} — it was not "
           "refreshed and no longer matches the values comparison.")
    if events is not None:
        events.on_log(msg)
    log.info("matrix: %s", msg)


def _settle_formulas_twin(compare_call, out_path, do_write, events=None,
                          source_paths=(), commit_guard=None):
    """Refresh the live-formulas twin when `do_write`, else clear a stale prior one
    (CMP-AUD-082). Called after EVERY successful values commit so a `(formulas)` sibling
    can never outlive the generation it was built for: a values-only refresh, an
    over-limit skip, an inputs-changed skip, or a failed formulas commit all clear it."""
    wrote = False
    if do_write:
        wrote = _try_formulas(compare_call, out_path, events,
                              source_paths=source_paths, commit_guard=commit_guard)
    if not wrote:
        _clear_stale_formulas_twin(out_path, events, commit_guard, source_paths)


# --------------------------------------------------------------------------- #
# CMP-AUD-098: pre-comparison input capture. The cache record must bind the
# output to the bytes the comparator actually READ — fingerprinting after
# production would bind a mid-comparison mutation's NEW identity to a workbook
# built from the OLD bytes and render the raced cell permanently fresh.
# --------------------------------------------------------------------------- #
def _fingerprint_for_record(fp_before, folders, out_name, events=None):
    """The input fingerprint to RECORD: always the PRE-comparison capture. When
    the folders' identity moved during the build, the recorded (pre) value no
    longer matches the current folders, so the freshness reader immediately
    reports the cell STALE — the raced result is invalidated, never fresh —
    and the race is announced (log + events)."""
    if _cell_input_fingerprint(*folders) != fp_before:
        msg = (f"sources changed while {out_name} was being built; the "
               "comparison is recorded already-stale and must be rebuilt")
        log.warning("matrix: %s", msg)
        if events is not None:
            events.on_log(f"⚠ {msg}")
    return fp_before


def _twin_inputs_unchanged(fp_before, folders, out_name, events=None):
    """True iff the source folders still match the pre-comparison capture; a
    changed identity SKIPS the live-formulas twin (it would be built from
    different bytes than the just-committed values workbook), loudly."""
    if _cell_input_fingerprint(*folders) == fp_before:
        return True
    msg = (f"sources changed after the values workbook of {out_name} was "
           "built; skipping the live-formulas copy (it would not match)")
    log.warning("matrix: %s", msg)
    if events is not None:
        events.on_log(f"⚠ {msg}")
    return False


# --------------------------------------------------------------------------- #
# orchestration: build one cell's comparison (pure delegation to compare_env)
# --------------------------------------------------------------------------- #
def build_cell_comparison(dest, baseline_key, row_key, cell_key, events,
                          confirm_overwrite=None, row_defs=None, also_formulas=False,
                          commit_guard=None):
    """Compare (row's report, cell env) against the baseline env, writing the
    VALUES workbook to comparison_path(...), and record its verdict + discrepancy
    counts in the cache. Pure delegation to the adapter's compare_folders — the
    comparison engine is untouched. Returns the ConsolidateResult. With
    `also_formulas`, also writes a live-formulas twin beside the values copy.

    Raises ValueError on an unknown row_key or a baseline cell (nothing to
    compare); compare_folders itself returns a clean error result when a side
    hasn't been exported yet."""
    rows = row_defs if row_defs is not None else _row_defs()
    if row_key not in rows:
        raise ValueError(f"unknown matrix row: {row_key}")
    if cell_key == baseline_key:
        raise ValueError("the baseline column has nothing to compare against")
    _label, subdir, _idx, adapter, _hr = rows[row_key]   # _hr: layout is detected from the workbook (O4)

    dest = Path(dest)
    out_path = comparison_path(dest, baseline_key, row_key, cell_key)
    _require_commit_guard(commit_guard, "comparison build", out_path)
    source_paths = (dest / cell_key, dest / baseline_key)
    # CMP-AUD-098: capture the input identity BEFORE the comparator reads.
    fp_folders = (dest / cell_key / subdir, dest / baseline_key / subdir)
    fp_before = _cell_input_fingerprint(*fp_folders)

    # side A = the cell env, side B = the baseline (labels read "ARS-TEST vs SSOR-PROD").
    # F9/S2: the public adapter owns its atomic temp/validation/publication
    # transaction and receives the Matrix's exact target-aware lease directly.
    result = adapter.compare_folders(
        dest / cell_key, dest / baseline_key, out_path, events=events,
        confirm_overwrite=confirm_overwrite or (lambda _p: True), mode="values",
        commit_guard=commit_guard)
    if result.status == "ok":
        # CMP-AUD-082: refresh the live-formulas twin, or clear a stale prior one.
        do_write = also_formulas and _twin_inputs_unchanged(
            fp_before, fp_folders, out_path.name, events)
        _settle_formulas_twin(lambda fp: adapter.compare_folders(
            dest / cell_key, dest / baseline_key, fp, events=events,
            confirm_overwrite=lambda _p: True, mode="formulas",
            commit_guard=commit_guard), out_path, do_write, events,
            source_paths=source_paths, commit_guard=commit_guard)

    if result.status == "ok" and out_path.exists():
        _require_commit_guard(commit_guard, "comparison cache write")
        # O4: detect the layout from the produced workbook, not the row's declared
        # has_route — an aggregate cross-env adapter (Ramp/Intersection Summary) can
        # have a sheet_name yet emit a flat sheet, which would otherwise read as 0 diffs.
        published = _m._published_comparison_result(out_path, result)
        typed = published.comparison_outcome
        diff_cells = typed.counts.differing_cells
        one_sided = (typed.counts.side_a_only_rows
                     + typed.counts.side_b_only_rows)
        try:
            built_at = out_path.stat().st_mtime
        except OSError:
            built_at = None
        # F5/P2: record the inputs' identity (cell + baseline export folders, in that
        # order) so a later snapshot reads the cell stale when a route changed.
        # CMP-AUD-098: the PRE-comparison capture is recorded — a mid-build
        # mutation therefore reads immediately stale, never fresh.
        record_result(dest, baseline_key, row_key, cell_key, typed.verdict,
                      diff_cells, one_sided, built_at,
                      completion=typed.completion,
                      input_fingerprint=_fingerprint_for_record(
                          fp_before, fp_folders, out_path.name, events),
                      generation_id=published.artifact_generation.generation_id,
                      producer_versions=producer_identity(),
                      commit_guard=commit_guard)
    return result


def cells_to_rebuild(snapshot, scope="stale", row=None, env=None):
    """[(row_key, cell_key, mode_id)] to (re)compute, honoring each row's SELECTED
    mode. scope='all' = every comparable cell (both sides present); 'stale' = only
    missing/stale ones. Optional `row` / `env` filters drive the per-row and
    per-column refresh buttons. Skips the env-mode baseline column, unsupported
    (greyed) modes, and cells with a missing input side."""
    modes = snapshot.get("modes", {})
    todo = []
    for row_key in snapshot["rows"]:
        if row and row_key != row:
            continue
        mode_id = modes.get(row_key, "env")
        for ev in snapshot["envs"]:
            if env and ev != env:
                continue
            cmp = snapshot["cells"][row_key][ev].get("cmp")
            if cmp is None:                      # env-mode baseline column
                continue
            if not cmp.get("supported") or cmp.get("missing_side"):
                continue
            if scope == "all" or cmp.get("stale"):
                todo.append((row_key, ev, mode_id))
    return todo


# --------------------------------------------------------------------------- #
# orchestration: consolidate the env's store folder(s) and diff via the existing
# (untouched) comparison adapters. compare_highway_log / compare_highway_log_pdf
# are file-vs-file, so the per-route env folders are consolidated first.
# --------------------------------------------------------------------------- #
def _pdf_store_consolidator(subdir):
    """The consolidator module for a PDF-sourced report — Highway Log (PDF) and
    Intersection Detail (PDF), the two reports deliberately absent from
    reports.consolidator_for_subdir because they need a scratch converted_dir to
    parse the PDFs first. Used by BOTH _consolidate_store_folder and
    _consolidated_filename so a PDF report is wired in ONE place — a matrix/by-day row
    whose subdir resolved no consolidated filename was the v0.17.4 by-day crash class."""
    if subdir == "highway_log_pdf":
        import consolidate_tsmis_highway_log_pdf as _m       # pdfplumber — lazy
        return _m
    if subdir == "intersection_detail_pdf":
        import consolidate_tsmis_intersection_detail_pdf as _m
        return _m
    if subdir == "highway_detail_pdf":
        import consolidate_tsmis_highway_detail_pdf as _m
        return _m
    if subdir == "highway_sequence_pdf":
        import consolidate_tsmis_highway_sequence_pdf as _m
        return _m
    if subdir == "ramp_detail_pdf":
        import consolidate_tsmis_ramp_detail_pdf as _m
        return _m
    return None


def _consolidate_store_folder(subdir, env_dir, out_path, events,
                              commit_guard=None):
    """Consolidate one Export-Everything store folder (<env>/<subdir>/, per-route
    files) into a single workbook via the report's existing consolidator (with its
    additive input_dir/out_path override). Registry-driven via
    reports.consolidator_for_subdir, so any consolidatable report works; the PDF
    Highway Log is the one special case (needs a scratch converted_dir).

    Returns the consolidator's ConsolidateResult (F3) so callers can honor its
    completion — its `status`/`completion` decides whether the output is safe to
    compare and cache as a fresh result."""
    out_path = Path(out_path)
    env_dir = Path(env_dir)
    _require_commit_guard(commit_guard, "store consolidation", out_path)
    _require_commit_guard(commit_guard, "store input read", env_dir)
    # P2-A02: capture the inputs' identity BEFORE the build, so write_consolidated_fingerprint
    # can refuse to certify the workbook "fresh" if an external writer changed the source folder
    # mid-build (the GUI task lock already serializes our own writers).
    fp_before = artifact_store.fingerprint(env_dir)
    pdf_mod = _pdf_store_consolidator(subdir)
    if pdf_mod is not None:
        # Highway Log (PDF) / Intersection Detail (PDF): parse the per-route PDFs into
        # a scratch converted_dir first, then combine — they have no entry in
        # consolidator_for_subdir for exactly this reason.
        _require_commit_guard(
            commit_guard, "PDF conversion scratch parent", out_path.parent)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        _require_commit_guard(
            commit_guard, "PDF conversion scratch parent", out_path.parent)
        conv = Path(tempfile.mkdtemp(
            prefix=f".{out_path.stem}_conv-", dir=out_path.parent))
        conv_identity = owned_dir.directory_identity(conv)

        def _pdf_guard(path):
            # Bind every conversion mutation to the exact OS-created scratch
            # directory as well as the worker's comparisons/store roots.
            return (conv_identity is not None
                    and owned_dir.directory_identity(conv) == conv_identity
                    and _guard_allows(commit_guard, path))

        if not _pdf_guard(conv):
            raise ValueError(
                "The PDF conversion scratch directory changed during creation; "
                "it was retained and the comparison was not built.")
        try:
            res = pdf_mod.consolidate(events=events, confirm_overwrite=lambda _p: True,
                                      input_dir=env_dir, out_path=out_path,
                                      converted_dir=conv,
                                      commit_guard=_pdf_guard)
        finally:
            # A replaced/linked scratch pathname or linked child is no longer
            # ours to traverse recursively.
            if (_pdf_guard(conv)
                    and owned_dir.is_plain_directory_tree(
                        conv, identity=conv_identity)):
                shutil.rmtree(conv, ignore_errors=True)
            else:
                log.error("matrix: retained unsafe PDF conversion scratch path %s", conv)
    else:
        import reports                               # lazy (avoid import cycle)
        mod = reports.consolidator_for_subdir(subdir)
        if mod is None:
            raise ValueError(f"no store consolidator for {subdir}")
        # F3: RETURN the ConsolidateResult so callers can honor its completion — a
        # failed / no-data consolidation must not be compared or cached as fresh.
        res = mod.consolidate(events=events, confirm_overwrite=lambda _p: True,
                              input_dir=env_dir, out_path=out_path,
                              commit_guard=commit_guard)
    # P1-R01: persist the producer completion beside the consolidated workbook through
    # the shared boundary so a later REUSE (no fresh ConsolidateResult) still knows it
    # was built from partial inputs. No-op unless an output was produced (status ok). A
    # False return = a non-complete artifact's flag could NOT be recorded (publication
    # failed): the operation cannot claim a safely persisted artifact, so raise — the
    # comparison surfaces not-refreshed and keeps the prior cell rather than diffing a
    # workbook that might read complete.
    if not consolidation_meta.write_outcome(
            out_path, res, extra=getattr(res, "producer_extra", None),
            commit_guard=commit_guard):
        raise ValueError(f"the {subdir} consolidation finished but its outcome could not "
                         "be recorded; the incomplete workbook was invalidated — re-run")
    # P2/F5: record the source folder's input fingerprint beside the workbook so a later
    # reuse rebuilds when the inputs' IDENTITY changed (a route added / removed / resized),
    # not just when the newest mtime advanced (_consolidated_stale via consolidated_fresh).
    # P2-A02: pass the pre-build fingerprint — if the inputs changed during the build, the
    # workbook is NOT certified fresh (it rebuilds next reuse) rather than stamped wrong.
    if getattr(res, "status", None) == "ok":
        artifact_store.write_consolidated_fingerprint(
            out_path, env_dir, built_from=fp_before,
            commit_guard=commit_guard)
    return res


def consolidate_tsn_pdfs(dest, subdir, events=None, confirm_overwrite=None):
    """Build a consolidated TSN workbook FROM the district PDFs the user dropped in
    <dest>/_tsn_input/<subdir>/, writing it back there so the next tsn_source()
    finds it. TSN is the SAME district-PDF set for both Highway Log flavors, so
    only the 'highway_log' TSN drop folder is handled. Returns the out path."""
    if subdir != "highway_log":
        raise ValueError(f"no TSN PDF consolidator for {subdir}")
    import consolidate_tsn_highway_log as _ctsn   # pdfplumber — lazy
    in_dir = tsn_input_root(dest, subdir)
    out_path = in_dir / "tsn_highway_log_consolidated.xlsx"
    res = _ctsn.consolidate(events=events, confirm_overwrite=confirm_overwrite or (lambda _p: True),
                            input_dir=in_dir, out_path=out_path)
    # P1-B05: honor the producer result. A failed / no-data / cancelled TSN consolidation
    # must NOT return a success-shaped path — its worker would log "TSN workbook ready"
    # with errors=0. Raise the consolidator's own message so the worker reports a
    # not-refreshed failure (errors>0) and the prior workbook is kept untouched.
    if not outcome.comparable(outcome.consolidate_completion_of(res)):
        raise ValueError(res.message or f"the {subdir} TSN consolidation did not complete")
    # A partial set (some district PDFs left out) stays usable but flagged — persist the
    # producer completion beside the workbook so the matrix flags it on reuse. If the
    # flag could not be recorded (publication failed), the partial workbook was
    # invalidated — raise so the worker reports a not-refreshed failure (never a
    # success-shaped path to a workbook that would read complete).
    if not consolidation_meta.write_outcome(out_path, res):
        raise ValueError(f"the {subdir} TSN consolidation finished but its outcome could "
                         "not be recorded; the incomplete workbook was discarded — re-run")
    return out_path


# --- persistent (reusable) consolidated workbooks ------------------------- #
# v0.16.x: instead of consolidating a per-route store folder to a throwaway temp
# on EVERY comparison, persist the consolidated workbook into the run/store
# folder's `consolidated/` dir — the SAME filename + location the Consolidate tab
# uses — and REUSE it until the per-route files change. So: re-exporting a report
# makes its consolidated stale (a source file is newer) → the next comparison
# re-consolidates; only changing the comparison mechanism (not the data) reuses
# the existing consolidated. A force flag rebuilds it on demand.
def _consolidated_filename(subdir):
    pdf_mod = _pdf_store_consolidator(subdir)
    if pdf_mod is not None:
        return pdf_mod.FILENAME
    mod = reports.consolidator_for_subdir(subdir)
    if mod is None:
        raise ValueError(f"no consolidated filename for {subdir}")
    return mod.FILENAME


def consolidated_store_path(store_dir, subdir):
    """The PERSISTENT consolidated workbook for a per-route store folder: a sibling
    `consolidated/` dir, date/env-stamped via the parent run-folder name (so a day
    folder gets a stamped name and matches the Consolidate-tab output; the always-
    current Everything store, whose parent is just `<src-env>`, gets the plain
    name)."""
    from paths import stamped_consolidated_filename     # lazy (avoid import cycle)
    parent = Path(store_dir).parent
    name = stamped_consolidated_filename(_consolidated_filename(subdir), parent.name)
    return parent / "consolidated" / name


def consolidated_state(store_dir, subdir):
    """{exists, fresh, path} for a store folder's persistent consolidated — drives
    the 'consolidated for this day' indicator. fresh = present AND its inputs' IDENTITY
    is unchanged since it was built (the artifact_store input fingerprint, not a
    newest-mtime check — see _consolidated_stale; P2/F5)."""
    p = consolidated_store_path(store_dir, subdir)
    record = consolidation_meta.read_outcome(p) if p.exists() else None
    identity_fresh = artifact_store.consolidated_fresh(p, store_dir)
    trusted = bool(record is not None and record.trusted and record.current)
    completion = record.completion if trusted else None
    return {
        "exists": p.exists(),
        "fresh": bool(identity_fresh and trusted
                      and completion == outcome.COMPLETE),
        "completion": completion,
        "trusted": trusted,
        "path": str(p),
    }


def _consolidated_stale(consolidated, store_dir):
    """True when the persistent consolidated must be REBUILT: missing/unreadable, OR its
    inputs' IDENTITY changed since it was built — a route file added / removed / resized /
    re-timed (F5/R1-R03, via artifact_store's input fingerprint). The old signal was the
    store's NEWEST mtime, which missed a DELETED non-newest route: the consolidated then
    read as fresh though it still carried the gone route's rows. A legacy consolidated
    with no fingerprint sidecar reads stale ONCE, rebuilds, and records the sidecar (the
    one-time migration). Never raises."""
    if not artifact_store.consolidated_fresh(consolidated, store_dir):
        return True
    # CMP-AUD-084: a parser / consolidator fix ships in a new app version, so a
    # workbook built by an OLDER pipeline (with unchanged raw inputs) must re-parse
    # once — otherwise a corrected comparator keeps reading pre-fix rows. The stamped
    # producer version must equal the current one; a legacy workbook with no stamp
    # (read_extra default None) reads stale and rebuilds once, then records the stamp.
    if (consolidation_meta.read_extra(consolidated, "producer_app_version")
            != producer_identity().get("app")):
        return True
    record = consolidation_meta.read_outcome(consolidated)
    return bool(record is None or not record.trusted or not record.current
                or not outcome.comparable(record.completion))


def evidence_opts_for(evidence, row_key, dir_for_subdir):
    """Resolve one cell's visual-evidence request: None unless the user toggle
    is on AND the row supports evidence. `dir_for_subdir` maps a TSMIS export
    subdir to THIS cell's folder for it (the Everything matrix and the by-day
    matrix lay their stores out differently — each caller passes its own
    mapping, the resolution logic lives once here)."""
    if not (evidence and evidence.get("enabled")):
        return None
    import visual_evidence                               # lazy: pulls PIL/pdfium
    if not visual_evidence.capable(row_key):
        return None
    return {"tsmis_pdf_dir": dir_for_subdir(visual_evidence.pdf_subdir_for(row_key)),
            "examples": visual_evidence.clamp_examples(evidence.get("examples"))}


def run_evidence_only(row_key, store_dir, subdir, tsn_path, comparison_path,
                      tsmis_pdf_dir, events, examples=None, commit_guard=None,
                      source_identity_check=None, expected_generation_id=None,
                      source_workbook_identity=None, live_tsn_path=None,
                      _captured_tsn=False):
    """Generate/refresh the evidence set for an EXISTING vs-TSN comparison — the
    on-demand per-cell action (the toggle-driven decoration runs inside
    consolidate_and_compare_tsn as the comparison is built). Runs regardless of
    the evidence toggle; `examples` is engine-clamped.

    The FRESHNESS GATE is what keeps the on-demand path honest: the generator
    re-enumerates the differences from the CURRENT consolidated + TSN artifacts,
    so if either moved on since the comparison was built, the images could
    illustrate a diff set the workbook doesn't carry. It therefore REFUSES —
    with a "refresh the comparison" hint — when the store changed under the
    consolidated workbook, or when the consolidated/TSN workbook is newer than
    the comparison. Returns a ConsolidateResult; raises ValueError with the
    actionable reason for every not-runnable case."""
    import visual_evidence                               # lazy: pulls PIL/pdfium
    if not visual_evidence.capable(row_key):
        raise ValueError("this report doesn't support evidence images")
    if not _captured_tsn:
        if source_workbook_identity is None:
            raise ValueError(
                "The TSN workbook has no resolved content identity and cannot be "
                "safely captured for evidence.")
        _require_source_identity(source_identity_check, "capturing evidence inputs")
        live_tsn_path = Path(live_tsn_path or tsn_path)
        with captured_tsn_workbook(tsn_path, source_workbook_identity) as captured:
            return run_evidence_only(
                row_key, store_dir, subdir, captured, comparison_path,
                tsmis_pdf_dir, events, examples=examples,
                commit_guard=commit_guard,
                source_identity_check=source_identity_check,
                expected_generation_id=expected_generation_id,
                source_workbook_identity=source_workbook_identity,
                live_tsn_path=live_tsn_path, _captured_tsn=True)
    comparison_path = Path(comparison_path)
    tsn_path = Path(tsn_path)
    live_tsn_path = Path(live_tsn_path or tsn_path)
    evidence_guard = _compose_source_guard(commit_guard, source_identity_check)
    _require_source_identity(source_identity_check, "generating evidence")
    _require_commit_guard(evidence_guard, "visual-evidence write", comparison_path)
    consolidated = consolidated_store_path(store_dir, subdir)
    if not comparison_path.exists():
        raise ValueError("no comparison workbook for this cell yet — run the "
                         "comparison first")
    comparison_record = consolidation_meta.read_comparison_outcome(comparison_path)
    if (comparison_record is None or not comparison_record.trusted
            or not comparison_record.current
            or comparison_record.comparison_outcome is None):
        raise ValueError(
            "the comparison outcome metadata is missing or untrusted — refresh "
            "the comparison before generating evidence")
    if (not isinstance(expected_generation_id, str)
            or comparison_record.artifact_generation.generation_id
            != expected_generation_id):
        raise ValueError(
            "the comparison workbook and its TSN cache generation do not match — "
            "refresh the comparison before generating evidence")
    if comparison_record.comparison_outcome.completion != outcome.COMPLETE:
        raise ValueError(_partial_comparison_reason(
            comparison_record.comparison_outcome))
    if not consolidated.exists():
        raise ValueError(f"no consolidated {subdir} workbook for this cell — "
                         "run the comparison first")
    refresh_hint = ("— refresh the comparison instead (it regenerates the "
                    "evidence set when the Evidence images option is on, or "
                    "run this action again after)")
    if _m._consolidated_stale(consolidated, store_dir):
        raise ValueError(f"the {subdir} exports changed since the comparison "
                         f"was built {refresh_hint}")
    cmp_mtime = _safe_mtime(comparison_path) or 0
    if (_safe_mtime(consolidated) or 0) > cmp_mtime + _MTIME_TOL_S:
        raise ValueError(f"the consolidated {subdir} workbook is newer than "
                         f"this comparison {refresh_hint}")
    if (_safe_mtime(live_tsn_path) or 0) > cmp_mtime + _MTIME_TOL_S:
        raise ValueError(f"the TSN workbook is newer than this comparison "
                         f"{refresh_hint}")
    ev = visual_evidence.generate(
        row_key, consolidated, tsn_path, comparison_path, tsmis_pdf_dir, events,
        examples=visual_evidence.clamp_examples(examples),
        commit_guard=evidence_guard)
    _require_source_identity(source_identity_check, "publishing evidence")
    after_record = consolidation_meta.read_comparison_outcome(comparison_path)
    if (after_record is None or not after_record.trusted
            or after_record.artifact_generation
            != comparison_record.artifact_generation):
        raise ValueError(
            "the comparison generation changed while evidence was rendering — "
            "refresh the comparison and evidence")
    note = ev.get("note") or "evidence run finished"
    return ConsolidateResult(status="ok", message=note, summary_lines=[note])


def evidence_for_cell(dest, row_key, cell_key, baseline_key, events,
                      tsn_files=None, examples=None, commit_guard=None):
    """On-demand evidence for one Everything-matrix cell's EXISTING vs-TSN
    comparison. Resolves the same paths build_comparison's tsn branch uses —
    but consolidates nothing, compares nothing, and does NOT heal the TSN
    library (a heal would rebuild it newer than the comparison and the
    freshness gate would then rightly refuse; a version-stale library needs a
    comparison refresh anyway)."""
    rows = _row_defs()
    if row_key not in rows:
        raise ValueError(f"unknown matrix row: {row_key}")
    _label, subdir, _idx, adapter, _hr = rows[row_key]
    mode = _mode_by_id(_row_modes(row_key, subdir, adapter), "tsn")
    if not mode["supported"]:
        raise ValueError(f"no TSN comparison for {row_key}")
    dest = Path(dest)
    import tsn_library
    tsn_files, _selection_map_changed = tsn_library.canonicalize_selections(
        tsn_files or {})
    tsn_key = tsn_library.canonical_dataset_key(mode["tsn_subdir"])
    src = tsn_source(dest, tsn_key, tsn_files.get(tsn_key))
    if src.get("kind") == "missing_explicit":
        raise ValueError(tsn_library.explicit_selection_problem(src))
    if src.get("kind") not in ("file", "consolidated"):
        raise ValueError("no consolidated TSN workbook available")
    token, source_identity_check = tsn_identity_check_for(tsn_key, src)
    source_workbook_identity = tsn_expected_workbook_identity(
        tsn_key, src, token)
    cached = _m.load_tsn_results(dest)
    record = cached.get(f"{row_key}|{mode['id']}", {}).get(cell_key)
    expected_generation_id = require_cached_tsn_identity(record, token)
    import visual_evidence                               # lazy: pulls PIL/pdfium
    result = run_evidence_only(
        row_key, dest / cell_key / mode["env_subdir"], mode["env_subdir"],
        src["path"], mode_out_path(dest, baseline_key, row_key, cell_key, mode),
        dest / cell_key / visual_evidence.pdf_subdir_for(row_key),
        events, examples=examples, commit_guard=commit_guard,
        source_identity_check=source_identity_check,
        expected_generation_id=expected_generation_id,
        source_workbook_identity=source_workbook_identity,
        live_tsn_path=src["path"])
    if src.get("selection"):
        tsn_library.require_explicit_selection(src["selection"])
    return result


# Producer-completion persistence for the persistent consolidated workbook lives in
# the shared `consolidation_meta` boundary (P1-R01) — every persistent writer (matrix
# store consolidation, auto-consolidate, the GUI/console Consolidate tab, TSN-library
# builds) records the outcome there, and reuse recovers it, so no writer can bypass it.
def consolidate_and_compare_tsn(tsmis_store_dir, tsn_path, out_path, row_key, subdir,
                                events, confirm_overwrite=None, force_consolidate=False,
                                also_formulas=False, evidence_opts=None,
                                explicit_selection=None, commit_guard=None,
                                source_identity_check=None,
                                source_workbook_identity=None,
                                _captured_tsn=False):
    """The SHARED TSN compare path used by BOTH matrices (the Everything matrix's
    latest-store cells AND the Compare tab's by-day cells).

    Consolidate a per-route TSMIS store folder (`tsmis_store_dir`, the `subdir`
    report) into its PERSISTENT `consolidated/` workbook (reused when still fresh;
    rebuilt when the inputs' IDENTITY changed — a route added/removed/resized, the
    artifact_store fingerprint — or `force_consolidate`), then compare it vs
    the consolidated TSN workbook with the row's comparator (`_m.tsn_comparator_for(
    row_key)` — Highway Log Excel/PDF, Ramp Detail, …), writing the VALUES workbook
    to `out_path`. Returns the ConsolidateResult. Pure delegation to the untouched
    consolidate_* / compare_* adapters; the matrices differ only by the source
    folder + the output path. (For HL the output is semantically identical to the prior
    fmt-keyed path — same consolidator + same comparator; the P2 atomic-write wrapper only
    changes the write mechanism, not the workbook content.)"""
    if not _captured_tsn:
        if source_workbook_identity is None:
            raise ValueError(
                "The TSN workbook has no resolved content identity and cannot be "
                "safely captured for comparison.")
        _require_source_identity(source_identity_check, "capturing the TSN workbook")
        with captured_tsn_workbook(tsn_path, source_workbook_identity) as captured:
            return consolidate_and_compare_tsn(
                tsmis_store_dir, captured, out_path, row_key, subdir, events,
                confirm_overwrite=confirm_overwrite,
                force_consolidate=force_consolidate,
                also_formulas=also_formulas, evidence_opts=evidence_opts,
                explicit_selection=explicit_selection, commit_guard=commit_guard,
                source_identity_check=source_identity_check,
                source_workbook_identity=source_workbook_identity,
                _captured_tsn=True)
    out_path = Path(out_path)
    _require_commit_guard(commit_guard, "TSN comparison build", out_path)
    cmp_mod = _m.tsn_comparator_for(row_key)
    if cmp_mod is None:
        raise ValueError(f"no TSN comparator for {row_key}")
    consolidated = consolidated_store_path(tsmis_store_dir, subdir)
    _require_commit_guard(commit_guard, "persistent consolidation", consolidated)
    cres = None
    if force_consolidate or _m._consolidated_stale(consolidated, tsmis_store_dir):
        # The shared Events callback can stop a producer at its poll points;
        # this is the explicit lease boundary before entering it.
        _require_commit_guard(commit_guard, "TSMIS consolidation", consolidated)
        cres = _m._consolidate_store_folder(
            subdir, Path(tsmis_store_dir), consolidated, events,
            commit_guard=commit_guard)
    # F3: honor the consolidation outcome instead of silently discarding it. A
    # failed / no-data / cancelled consolidation must NOT feed a comparison or be
    # cached as fresh — raise the consolidator's OWN message (clearer than the
    # generic backstop below), so the cell surfaces exactly why it didn't refresh
    # and the prior cached comparison is left untouched (not overwritten with a
    # result built from incomplete inputs). A partial consolidation still compares.
    # cres is None when the existing consolidated was fresh and reused.
    if cres is not None and not outcome.comparable(outcome.consolidate_completion_of(cres)):
        raise ValueError(cres.message or f"the {subdir} consolidation did not complete")
    # Backstop: a consolidator can return without raising yet write nothing (an
    # empty store folder). Catch that here so the failure names the consolidation
    # step instead of the compare adapter erroring on a missing/empty input.
    if not consolidated.exists() or consolidated.stat().st_size == 0:
        raise ValueError(f"nothing to compare — no {subdir} export found in "
                         f"{tsmis_store_dir}")
    # F9/S2: the public comparator owns the atomic transaction; the same target-aware
    # lease reaches its artifact temp and compare_core's exact save boundary.
    source_paths = (consolidated, tsn_path)
    comparison_guard = _compose_source_guard(
        commit_guard, source_identity_check)

    def _compare_checked(target, mode, confirm=None):
        _require_source_identity(source_identity_check, "running the comparison")
        if explicit_selection:
            import tsn_library
            tsn_library.require_explicit_selection(explicit_selection)
        compared = cmp_mod.compare(
            consolidated, str(tsn_path), target, events=events,
            confirm_overwrite=confirm or (lambda _p: True), mode=mode,
            commit_guard=comparison_guard)
        if explicit_selection:
            tsn_library.require_explicit_selection(explicit_selection)
        _require_source_identity(source_identity_check, "publishing the comparison")
        return compared

    # The direct comparator performs its own source-alias checks and atomic commit.
    result = _compare_checked(
        out_path, "values", confirm_overwrite or (lambda _p: True))
    if result.status == "ok":
        # CMP-AUD-082: refresh the live-formulas twin, or clear a stale prior one.
        _settle_formulas_twin(lambda fp: _compare_checked(fp, "formulas"), out_path,
            also_formulas, events, source_paths=source_paths,
            commit_guard=comparison_guard)
    # P1-R01: a PARTIAL TSMIS consolidation (inputs left out) still compares, but the
    # diff is over INCOMPLETE inputs — flag the comparison so the caller records the
    # cell as partial. `cres` is set only when we (re)consolidated THIS call; when the
    # existing consolidated was REUSED (`cres` is None) recover the producer completion
    # from the sidecar so a reused partial stays flagged instead of silently reading
    # as a fresh, complete cell.
    # compare_tsn_common reduces the persisted input outcomes before run_compare
    # publishes its typed result. Do not mutate completion after publication.
    # Visual evidence is a DECORATION of a finished comparison: it renders
    # highlighted PDF snippets for sampled diffs next to the workbook. It never
    # changes the comparison's status/completion/counts, and any failure only
    # logs + notes (the comparison already succeeded). `evidence_opts` is set by
    # the callers when the user's toggle is on AND the row supports it.
    if result.status == "ok" and evidence_opts:
        try:
            _require_commit_guard(comparison_guard, "visual-evidence write")
            published = consolidation_meta.require_published_comparison(
                out_path, result)
            if published.comparison_outcome.completion != outcome.COMPLETE:
                raise ValueError(_partial_comparison_reason(
                    published.comparison_outcome))
            if explicit_selection:
                import tsn_library
                tsn_library.require_explicit_selection(explicit_selection)
            import visual_evidence                       # lazy: pulls PIL/pdfium
            ev = visual_evidence.generate(
                row_key, consolidated, tsn_path, out_path,
                evidence_opts["tsmis_pdf_dir"], events,
                examples=evidence_opts.get("examples"),
                commit_guard=comparison_guard)
            if explicit_selection:
                tsn_library.require_explicit_selection(explicit_selection)
            _require_source_identity(source_identity_check, "publishing evidence")
            if ev.get("note"):
                result.summary_lines = list(result.summary_lines) + [ev["note"]]
        except Exception as e:                           # noqa: BLE001
            log.warning("evidence generation for %s skipped", row_key, exc_info=True)
            msg = str(e).splitlines()[0] if str(e) else type(e).__name__
            events.on_log(f"  evidence images skipped — {msg}")
            result.summary_lines = list(result.summary_lines) + [
                f"evidence images skipped — {msg}"]
    return result


def _ensure_consolidated(store_dir, subdir, events, force,
                         commit_guard=None):
    """Return (persistent consolidated workbook path, its producer completion) for a
    store folder, rebuilding when stale or forced. Shared by the self (PDF-vs-Excel)
    path. F3: raises a clear ValueError when a rebuild did not complete (failed/
    no-data/cancelled), so the self comparison names the consolidation failure instead
    of erroring later on a missing/stale input. P1-R01: the completion is the fresh
    build's (when rebuilt) or the persisted sidecar's (when reused), so a partial side
    stays flagged through the self comparison."""
    p = consolidated_store_path(store_dir, subdir)
    _require_commit_guard(commit_guard, "persistent consolidation", p)
    comp = None
    if force or _m._consolidated_stale(p, store_dir):
        cres = _m._consolidate_store_folder(
            subdir, Path(store_dir), p, events,
            commit_guard=commit_guard)
        if cres is not None and not outcome.comparable(outcome.consolidate_completion_of(cres)):
            raise ValueError(cres.message or f"the {subdir} consolidation did not complete")
        comp = outcome.consolidate_completion_of(cres) if cres is not None else None
    if comp is None:                       # reused (or rebuild gave no result) -> sidecar
        record = consolidation_meta.read_outcome(p)
        if record is None or not record.trusted or not record.current:
            raise ValueError(
                f"the {subdir} consolidation outcome is missing or untrusted; rebuild it")
        comp = record.completion
    return p, comp


def build_comparison(dest, row_key, cell_key, mode_id, baseline_key, events,
                     tsn_files=None, confirm_overwrite=None, row_defs=None,
                     force_consolidate=False, also_formulas=False, evidence=None,
                     commit_guard=None):
    """(Re)build one cell's comparison for the row's SELECTED mode, write the VALUES
    workbook, and cache its counts. Dispatches to the existing comparison adapters
    (never edits them). Returns the ConsolidateResult. Raises ValueError for an
    unknown row or an unsupported/greyed mode; an absent input side yields the
    adapter's clean error result. `force_consolidate` rebuilds the persistent
    consolidated even when it looks fresh; `also_formulas` writes a live-formulas
    twin beside the values copy. `evidence` ({'enabled','examples'}) requests the
    visual-evidence decoration on a TSN-mode cell of a supported row."""
    rows = row_defs if row_defs is not None else _row_defs()
    if row_key not in rows:
        raise ValueError(f"unknown matrix row: {row_key}")
    _label, subdir, _idx, adapter, _hr = rows[row_key]
    mode = _mode_by_id(_row_modes(row_key, subdir, adapter), mode_id)
    if not mode["supported"]:
        raise ValueError(f"no comparison for {row_key} / {mode['id']}")
    _require_commit_guard(commit_guard, "matrix cell build")

    if mode["kind"] == "env":
        return _m.build_cell_comparison(dest, baseline_key, row_key, cell_key, events,
                                     confirm_overwrite=confirm_overwrite, row_defs=rows,
                                     also_formulas=also_formulas,
                                     commit_guard=commit_guard)

    dest = Path(dest)
    out_path = mode_out_path(dest, baseline_key, row_key, cell_key, mode)
    # CMP-AUD-098: capture the TSMIS source-folder identity BEFORE any
    # consolidation or comparison work reads it (the automatic
    # consolidate→compare chain included); same folders/order as the snapshot.
    if mode["kind"] == "tsn":
        fp_folders = (dest / cell_key / mode["env_subdir"],)
    else:                                # self: TSMIS PDF vs Excel (both folders)
        fp_folders = (dest / cell_key / mode["env_subdir"],
                      dest / cell_key / mode["other_subdir"])
    fp_before = _cell_input_fingerprint(*fp_folders)
    import tsn_library                              # lazy: explicit-selection contract
    tsn_files, _selection_map_changed = tsn_library.canonicalize_selections(
        tsn_files or {})
    if mode["kind"] == "tsn":
        tsn_key = tsn_library.canonical_dataset_key(mode["tsn_subdir"])
        src = tsn_source(dest, tsn_key, tsn_files.get(tsn_key))
        if src.get("kind") == "missing_explicit":
            raise ValueError(tsn_library.explicit_selection_problem(src))
        if src.get("kind") not in ("file", "consolidated"):
            raise ValueError("no consolidated TSN workbook available")
        if src.get("kind") == "consolidated":
            # D2/CMP-AUD-035: ensure_current either certifies/rebuilds this
            # canonical artifact or returns a typed error. In particular, a
            # stale workbook whose raw is gone/unreadable/ambiguous must stop
            # here and can never reach consolidate_and_compare_tsn.
            healed = tsn_library.ensure_current(tsn_key, events, source=src)
            if healed is not None:
                if healed.status != "ok":
                    raise ValueError(healed.message
                                     or "the TSN library is not certifiably current; comparison stopped")
            # Re-resolve even when ensure_current is a no-op: this is the exact
            # content token that the comparison and its cache must bind.
            src = tsn_source(dest, tsn_key, tsn_files.get(tsn_key))
            if src.get("kind") != "consolidated":
                raise ValueError(
                    "the canonical TSN source changed while it was being certified")
        tsn_token, source_identity_check = tsn_identity_check_for(tsn_key, src)
        source_workbook_identity = tsn_expected_workbook_identity(
            tsn_key, src, tsn_token)
        result = _m.consolidate_and_compare_tsn(
            dest / cell_key / mode["env_subdir"], src["path"], out_path,
            row_key, mode["env_subdir"], events, confirm_overwrite=confirm_overwrite,
            force_consolidate=force_consolidate, also_formulas=also_formulas,
            evidence_opts=evidence_opts_for(evidence, row_key,
                                            lambda sub: dest / cell_key / sub),
            explicit_selection=src.get("selection"),
            commit_guard=commit_guard,
            source_identity_check=source_identity_check,
            source_workbook_identity=source_workbook_identity)
        # P1-B05: the TSN side can ALSO be partial (a TSN builder left categories /
        # district PDFs out). consolidate_and_compare_tsn reduced the TSMIS side; reduce
        # the TSN side here (its completion rides on the resolved source) so the cell
        # flags partial when EITHER input was incomplete.
    else:                                # self: TSMIS PDF vs Excel (persisted both sides)
        side_env, comp_env = _m._ensure_consolidated(
            dest / cell_key / mode["env_subdir"], mode["env_subdir"], events,
            force_consolidate, commit_guard=commit_guard)
        side_other, comp_other = _m._ensure_consolidated(
            dest / cell_key / mode["other_subdir"], mode["other_subdir"], events,
            force_consolidate, commit_guard=commit_guard)
        # the adapter fixes PDF=side A, Excel=side B regardless of the row. The PDF
        # side is whichever subdir ends with `_pdf` (env on the …_pdf row, other on the
        # Excel row); the comparator is that family's PDF-vs-Excel adapter.
        env_is_pdf = mode["env_subdir"].endswith("_pdf")
        pdf_subdir = mode["env_subdir"] if env_is_pdf else mode["other_subdir"]
        pdf_c = side_env if env_is_pdf else side_other
        excel_c = side_other if env_is_pdf else side_env
        _self_cmp = _pdf_self_comparator(pdf_subdir)
        source_paths = (pdf_c, excel_c)
        result = _self_cmp.compare(
            pdf_c, excel_c, out_path, events=events,
            confirm_overwrite=confirm_overwrite or (lambda _p: True),
            mode="values", commit_guard=commit_guard)
        if result.status == "ok":
            # CMP-AUD-082: refresh the live-formulas twin, or clear a stale prior one.
            _settle_formulas_twin(lambda fp: _self_cmp.compare(
                pdf_c, excel_c, fp, events=events,
                confirm_overwrite=lambda _p: True, mode="formulas",
                commit_guard=commit_guard), out_path, also_formulas, events,
                source_paths=source_paths, commit_guard=commit_guard)
        # P1-R01: a PARTIAL side means the self comparison diffed INCOMPLETE inputs —
        # propagate it (either side partial -> the cell records partial) so a self
        # comparison can't hide that one side was built from incomplete inputs.

    if result.status == "ok" and out_path.exists():
        if mode["kind"] == "tsn":
            _require_source_identity(source_identity_check,
                                     "recording the comparison cache")
            cache_guard = _compose_source_guard(
                commit_guard, source_identity_check)
        else:
            cache_guard = commit_guard
        _require_commit_guard(cache_guard, "comparison cache write")
        # F4: detect the layout from the produced workbook. This path serves EVERY
        # report's vs-TSN/self comparison — Highway Log modes are Route-keyed, but
        # the aggregate reports (Ramp / Intersection Summary) emit a FLAT sheet, so
        # a hardcoded has_route=True read their Status/Diffs one column off and
        # reported 0 diffs. The header tells us the real layout regardless of mode.
        published = _m._published_comparison_result(out_path, result)
        typed = published.comparison_outcome
        diff_cells = typed.counts.differing_cells
        one_sided = (typed.counts.side_a_only_rows
                     + typed.counts.side_b_only_rows)
        # F5/P2: the cell's TSMIS source-folder identity, in the SAME order the
        # snapshot's _cmp_state fingerprints (env first; + 'other' for the self
        # PDF-vs-Excel mode). CMP-AUD-098: the PRE-comparison capture is what
        # gets recorded — a mid-build mutation reads immediately stale.
        record_tsn_result(dest, f"{row_key}|{mode['id']}", cell_key, typed.verdict,
                          diff_cells, one_sided, _safe_mtime(out_path),
                          completion=typed.completion,
                          input_fingerprint=_fingerprint_for_record(
                              fp_before, fp_folders, out_path.name, events),
                          source_identities=(
                              {"tsn": tsn_token}
                              if mode["kind"] == "tsn"
                              else {}),
                          generation_id=published.artifact_generation.generation_id,
                          producer_versions=producer_identity(),
                          commit_guard=cache_guard)
    return result
