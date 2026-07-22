"""Read a PUBLISHED comparison workbook's own cells (CMP-AUD-208/209/108).

Anything that claims to VERIFY a comparison has to consume what that comparison
PUBLISHED — not re-run the projection that produced it. Re-executing the
loaders answers "does the code agree with itself?"; a consistently wrong loader
passes that test twice. This module answers the other question: what does the
committed workbook actually say about each cell?

The values workbook carries that truth structurally:

  * ``Status``/``Diffs`` — the anchored count contract (shared with
    ``artifact_store.comparison_counts``; located by UNIQUE EXACT LABEL);
  * the hidden versioned ``__CMP_E1_STATE_V1_C###_P####_P####`` chunks — one
    character per displayed field, in display order, with the same positional
    meaning in both workbook twins: ``E`` equal, ``D`` different, ``N``
    displayed-but-non-asserting context/ditto, ``U`` one-sided;
  * the hidden versioned ``__CMP_E2_KEY_V1_TOKEN`` — the injective per-row
    token that also appears in each side's literal "Key (helper)" column, so a
    row's two SOURCE rows are recoverable without trusting Comparison's own
    hyperlinks (the CMP-AUD-218 mechanism).

Discrepancy state is read from the masks ONLY. The visible ``  ≠  `` separator
is content/presentation and is never scanned for state (CLAUDE.md); published
display text is compared for EQUALITY against the engine's own composition,
which is a different thing from inferring state from it.

Console-free; openpyxl is imported lazily so this module stays off GUI startup.
Raises ``PublishedComparisonError`` for anything it cannot read or authenticate
— a caller that cannot decode the published cells must not claim to verify
them.
"""
import hashlib
import json
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field as _dc_field
from pathlib import Path

log = logging.getLogger("tsmis.published")

# The reader's own version. Bump when the decoded shape changes meaning; the
# evidence workbook records it beside the ledger digest.
READER_VERSION = 1

STATE_EQUAL = "E"
STATE_DIFFERENT = "D"
STATE_CONTEXT = "N"
STATE_ONE_SIDED = "U"
STATES = (STATE_EQUAL, STATE_DIFFERENT, STATE_CONTEXT, STATE_ONE_SIDED)

STATUS_BOTH = "Both"
_SHEET = "Comparison"
_KEY_HELPER_HEADER = "Key (helper)"

_STATE_HEADER = re.compile(
    r"^__(?P<version>CMP_E\d+_STATE_V\d+)"
    r"_C(?P<chunk>\d{3})_P(?P<start>\d{4})_P(?P<end>\d{4})$")
_TOKEN_HEADER = re.compile(r"^__(?P<version>CMP_E\d+_KEY_V\d+)_TOKEN$")

# Comparison's id-header block is fixed by the layout:
#   [Route?] key "#" "<A> Row" "<B> Row" "Status" "Diffs"
# so the Status/Diffs anchor pair positions every other identity column.
_ID_TAIL = 2                       # Status, Diffs
_ID_HEAD_WITH_ROUTE = 5            # Route, key, #, A Row, B Row
_ID_HEAD_NO_ROUTE = 4              # key, #, A Row, B Row


class PublishedComparisonError(ValueError):
    """The published comparison could not be read or authenticated."""


@dataclass(frozen=True)
class PublishedRow:
    """One decoded Comparison row: its address, its state, its own values."""

    excel_row: int
    route: str
    key: str
    occurrence: int
    status: str
    diffs: int
    mask: str
    values: tuple
    token: str

    @property
    def address(self):
        """The published row's identity address (route, key, occurrence)."""
        return (self.route, self.key, self.occurrence)

    @property
    def matched(self):
        return self.status == STATUS_BOTH

    def state(self, position):
        return self.mask[position]

    def value(self, position):
        return self.values[position]


@dataclass(frozen=True)
class FieldLedger:
    """One displayed column's complete published accounting."""

    field: str
    position: int
    differences: int = 0
    solo_differences: int = 0          # on a key with exactly one matched row
    duplicate_differences: int = 0     # inside a repeated-key group
    context_cells: int = 0
    equal_cells: int = 0
    one_sided_cells: int = 0


@dataclass(frozen=True)
class ComparisonLedger:
    """The EXHAUSTIVE published universe, built before any sampling.

    CMP-AUD-209: no discrepancy class may be excluded before a display example
    is chosen. Every counted cell, every one-sided row, every duplicate group,
    and every context cell is accounted for here first, and the whole ledger is
    hash-bound so a later sample cannot silently narrow it.
    """

    fields: tuple
    data_rows: int
    matched_rows: int
    one_sided_rows: int
    one_sided_by_status: tuple
    duplicate_groups: int
    duplicate_member_rows: int
    difference_cells: int
    context_cells: int
    equal_cells: int
    one_sided_cells: int
    reader_version: int = READER_VERSION
    _by_field: dict = _dc_field(default_factory=dict, repr=False, compare=False)

    def for_field(self, name):
        return self._by_field.get(name)

    def differences(self, name):
        entry = self._by_field.get(name)
        return entry.differences if entry else 0

    def solo_differences(self, name):
        entry = self._by_field.get(name)
        return entry.solo_differences if entry else 0

    def duplicate_differences(self, name):
        entry = self._by_field.get(name)
        return entry.duplicate_differences if entry else 0

    def fields_with_differences(self):
        return tuple(entry.field for entry in self.fields if entry.differences)

    def as_payload(self):
        """The canonical serialization the digest is taken over."""
        return {
            "reader_version": self.reader_version,
            "data_rows": self.data_rows,
            "matched_rows": self.matched_rows,
            "one_sided_rows": self.one_sided_rows,
            "one_sided_by_status": [list(item)
                                    for item in self.one_sided_by_status],
            "duplicate_groups": self.duplicate_groups,
            "duplicate_member_rows": self.duplicate_member_rows,
            "difference_cells": self.difference_cells,
            "context_cells": self.context_cells,
            "equal_cells": self.equal_cells,
            "one_sided_cells": self.one_sided_cells,
            "fields": [
                {"field": entry.field, "position": entry.position,
                 "differences": entry.differences,
                 "solo_differences": entry.solo_differences,
                 "duplicate_differences": entry.duplicate_differences,
                 "context_cells": entry.context_cells,
                 "equal_cells": entry.equal_cells,
                 "one_sided_cells": entry.one_sided_cells}
                for entry in self.fields],
        }

    def digest(self):
        """Hash-bound completeness: the ledger's exact contents (CMP-AUD-209)."""
        blob = json.dumps(self.as_payload(), sort_keys=True,
                          separators=(",", ":"), ensure_ascii=False)
        return hashlib.sha256(blob.encode("utf-8")).hexdigest()


class PublishedComparison:
    """The decoded, authenticated cells of one committed comparison workbook."""

    def __init__(self, path, fields, rows, has_route, state_version,
                 token_version, side_labels):
        self.path = Path(path)
        self.fields = tuple(fields)
        self.rows = tuple(rows)
        self.has_route = has_route
        self.state_version = state_version
        self.token_version = token_version
        self.side_labels = tuple(side_labels)
        self._position = {name: i for i, name in enumerate(self.fields)}
        self._by_address = {row.address: row for row in self.rows}
        self._group_size = Counter((row.route, row.key) for row in self.rows)

    # -- addressing ------------------------------------------------------- #
    def position_of(self, name):
        try:
            return self._position[name]
        except KeyError:
            raise PublishedComparisonError(
                f"the published comparison has no column {name!r}") from None

    def require_fields(self, names):
        """Every name must be a published column, so a mask position can never
        be read against the wrong column (CMP-AUD-208)."""
        missing = [name for name in names if name not in self._position]
        if missing:
            raise PublishedComparisonError(
                f"the published comparison does not carry column(s) "
                f"{', '.join(missing)}; refresh the comparison")

    def row_at(self, route, key, occurrence=1):
        return self._by_address.get((route, key, occurrence))

    def group_size(self, route, key):
        return self._group_size[(route, key)]

    def is_solo(self, row):
        return self._group_size[(row.route, row.key)] == 1

    # -- the exhaustive ledger (CMP-AUD-209) ------------------------------ #
    def ledger(self):
        counts = {name: Counter() for name in self.fields}
        matched = one_sided = 0
        by_status = Counter()
        for row in self.rows:
            solo = self.is_solo(row)
            if row.matched:
                matched += 1
            else:
                one_sided += 1
                by_status[row.status] += 1
            for position, code in enumerate(row.mask):
                bucket = counts[self.fields[position]]
                bucket[code] += 1
                if code == STATE_DIFFERENT:
                    bucket["solo" if solo else "dup"] += 1
        entries = tuple(
            FieldLedger(field=name, position=position,
                        differences=counts[name][STATE_DIFFERENT],
                        solo_differences=counts[name]["solo"],
                        duplicate_differences=counts[name]["dup"],
                        context_cells=counts[name][STATE_CONTEXT],
                        equal_cells=counts[name][STATE_EQUAL],
                        one_sided_cells=counts[name][STATE_ONE_SIDED])
            for position, name in enumerate(self.fields))
        groups = [size for size in self._group_size.values() if size > 1]
        return ComparisonLedger(
            fields=entries,
            data_rows=len(self.rows),
            matched_rows=matched,
            one_sided_rows=one_sided,
            one_sided_by_status=tuple(sorted(by_status.items())),
            duplicate_groups=len(groups),
            duplicate_member_rows=sum(groups),
            difference_cells=sum(e.differences for e in entries),
            context_cells=sum(e.context_cells for e in entries),
            equal_cells=sum(e.equal_cells for e in entries),
            one_sided_cells=sum(e.one_sided_cells for e in entries),
            _by_field={e.field: e for e in entries})

    # -- persisted source rows (CMP-AUD-208) ------------------------------ #
    def source_rows(self, tokens):
        """``{token: {side_label: data-sheet row}}`` for the given row tokens.

        Resolved through each side's literal "Key (helper)" column — the same
        opaque token Spot Check MATCHes — so an evidence item names real
        persisted source rows without trusting Comparison's hyperlinks (which
        carry no cached value in a values workbook and read as blank).
        """
        wanted = {token for token in tokens if token}
        if not wanted:
            return {}
        found = defaultdict(dict)
        for side in self.side_labels:
            for token, row_number in _scan_helper_tokens(
                    self.path, side, wanted).items():
                found[token][side] = row_number
        return dict(found)


# --------------------------------------------------------------------------- #
# reading
# --------------------------------------------------------------------------- #
def _open(path):
    try:
        from openpyxl import load_workbook
    except ImportError as e:                     # pragma: no cover - deps gate
        raise PublishedComparisonError(
            "Required components are missing (openpyxl).") from e
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as e:                       # noqa: BLE001
        raise PublishedComparisonError(
            f"the published comparison could not be opened "
            f"({type(e).__name__}: {e})") from e


def _unique_column(header, label):
    hits = [i for i, value in enumerate(header) if value == label]
    if len(hits) != 1:
        raise PublishedComparisonError(
            f"the Comparison sheet has {len(hits)} {label!r} columns; "
            "exactly one is required")
    return hits[0]


def _state_columns(header):
    chunks = []
    for index, value in enumerate(header):
        if not isinstance(value, str):
            continue
        match = _STATE_HEADER.match(value)
        if match:
            chunks.append((index, match))
    if not chunks:
        raise PublishedComparisonError(
            "the Comparison sheet carries no state-mask chunks; it was not "
            "written by a current comparison")
    versions = {match.group("version") for _index, match in chunks}
    if len(versions) != 1:
        raise PublishedComparisonError(
            f"the Comparison sheet mixes state-mask versions {sorted(versions)}")
    expected_start = 0
    for position, (index, match) in enumerate(chunks, start=1):
        if int(match.group("chunk")) != position:
            raise PublishedComparisonError(
                "the Comparison state-mask chunks are out of order")
        if int(match.group("start")) != expected_start:
            raise PublishedComparisonError(
                "the Comparison state-mask chunks are not contiguous")
        expected_start = int(match.group("end")) + 1
    return [index for index, _match in chunks], versions.pop(), expected_start


def _token_column(header):
    hits = [(i, _TOKEN_HEADER.match(v)) for i, v in enumerate(header)
            if isinstance(v, str)]
    hits = [(i, m) for i, m in hits if m]
    if len(hits) != 1:
        raise PublishedComparisonError(
            "the Comparison sheet needs exactly one hidden row-token column")
    return hits[0][0], hits[0][1].group("version")


def _layout(header):
    """Locate every identity column from the anchored Status/Diffs pair."""
    status = _unique_column(header, "Status")
    diffs = _unique_column(header, "Diffs")
    if diffs != status + 1:
        raise PublishedComparisonError(
            "the Comparison sheet's Status/Diffs columns are not adjacent")
    if status == _ID_HEAD_WITH_ROUTE:
        has_route = True
    elif status == _ID_HEAD_NO_ROUTE:
        has_route = False
    else:
        raise PublishedComparisonError(
            f"the Comparison sheet has an unknown identity block "
            f"({status + _ID_TAIL} columns before the fields)")
    key_col = 1 if has_route else 0
    occ_col = key_col + 1
    side_a = header[occ_col + 1]
    side_b = header[occ_col + 2]
    labels = []
    for label in (side_a, side_b):
        if not isinstance(label, str) or not label.endswith(" Row"):
            raise PublishedComparisonError(
                "the Comparison sheet's source-row columns are unlabelled")
        labels.append(label[:-len(" Row")])
    return dict(status=status, diffs=diffs, has_route=has_route,
                key=key_col, occ=occ_col, sides=tuple(labels))


def _cell_text(value):
    return "" if value is None else str(value)


def _decode_row(excel_row, raw, lay, state_cols, token_col, n_fields):
    mask = "".join(_cell_text(raw[c]) if c < len(raw) else ""
                   for c in state_cols)
    if len(mask) != n_fields:
        raise PublishedComparisonError(
            f"Comparison row {excel_row} carries {len(mask)} state codes; "
            f"{n_fields} were expected")
    bad = sorted(set(mask) - set(STATES))
    if bad:
        raise PublishedComparisonError(
            f"Comparison row {excel_row} has unknown state code(s) {bad}")
    status = raw[lay["status"]]
    if not isinstance(status, str) or not status:
        raise PublishedComparisonError(
            f"Comparison row {excel_row} has no Status")
    diffs = raw[lay["diffs"]]
    if status == STATUS_BOTH:
        if (isinstance(diffs, bool) or not isinstance(diffs, (int, float))
                or not float(diffs).is_integer() or diffs < 0):
            raise PublishedComparisonError(
                f"Comparison row {excel_row} has an invalid Diffs value "
                f"{diffs!r}")
        diffs = int(diffs)
        if diffs != mask.count(STATE_DIFFERENT):
            raise PublishedComparisonError(
                f"Comparison row {excel_row} claims {diffs} difference(s) but "
                f"its state mask holds {mask.count(STATE_DIFFERENT)}")
        if STATE_ONE_SIDED in mask:
            raise PublishedComparisonError(
                f"Comparison row {excel_row} is matched but carries one-sided "
                "state codes")
    else:
        if diffs not in (None, ""):
            raise PublishedComparisonError(
                f"Comparison row {excel_row} is one-sided but carries "
                f"Diffs {diffs!r}")
        if set(mask) != {STATE_ONE_SIDED}:
            raise PublishedComparisonError(
                f"Comparison row {excel_row} is one-sided but its state mask "
                "is not entirely one-sided")
        diffs = 0
    occurrence = raw[lay["occ"]]
    if isinstance(occurrence, float) and occurrence.is_integer():
        occurrence = int(occurrence)
    if not isinstance(occurrence, int) or isinstance(occurrence, bool):
        raise PublishedComparisonError(
            f"Comparison row {excel_row} has a non-integer occurrence "
            f"{occurrence!r}")
    first_field = lay["diffs"] + 1
    values = tuple(_cell_text(raw[first_field + i]) if first_field + i < len(raw)
                   else "" for i in range(n_fields))
    return PublishedRow(
        excel_row=excel_row,
        route=_cell_text(raw[0]) if lay["has_route"] else "",
        key=_cell_text(raw[lay["key"]]),
        occurrence=occurrence,
        status=status,
        diffs=diffs,
        mask=mask,
        values=values,
        token=_cell_text(raw[token_col]))


def read(path, *, expect_fields=None, is_cancelled=None):
    """Decode and authenticate one published comparison's Comparison sheet.

    ``expect_fields`` — when given, the decoded display columns must equal it
    exactly, so a caller can never map a mask position onto the wrong column.
    """
    path = Path(path)
    workbook = _open(path)
    try:
        if _SHEET not in workbook.sheetnames:
            raise PublishedComparisonError(
                f"the workbook has no {_SHEET!r} sheet")
        sheet = workbook[_SHEET]
        stream = sheet.iter_rows(values_only=True)
        header = list(next(stream, ()) or ())
        if not header:
            raise PublishedComparisonError("the Comparison sheet is empty")
        lay = _layout(header)
        state_cols, state_version, n_fields = _state_columns(header)
        token_col, token_version = _token_column(header)
        fields = [_cell_text(v) for v in header[lay["diffs"] + 1:state_cols[0]]]
        if len(fields) != n_fields:
            raise PublishedComparisonError(
                f"the Comparison sheet shows {len(fields)} field column(s) but "
                f"its state mask covers {n_fields}")
        if expect_fields is not None and list(expect_fields) != fields:
            raise PublishedComparisonError(
                "the published comparison's columns are not the ones this "
                "report compares; refresh the comparison")
        rows = []
        tokens = set()
        for excel_row, raw in enumerate(stream, start=2):
            if raw is None or all(v is None for v in raw):
                continue
            if is_cancelled is not None and len(rows) % 5000 == 0 \
                    and is_cancelled():
                raise PublishedComparisonError(
                    "reading the published comparison was cancelled")
            row = _decode_row(excel_row, raw, lay, state_cols, token_col,
                              n_fields)
            if row.token in tokens:
                raise PublishedComparisonError(
                    f"Comparison row {excel_row} repeats row token "
                    f"{row.token!r}")
            tokens.add(row.token)
            rows.append(row)
        if not rows:
            raise PublishedComparisonError(
                "the published comparison holds no rows")
        return PublishedComparison(path, fields, rows, lay["has_route"],
                                   state_version, token_version, lay["sides"])
    finally:
        try:
            workbook.close()
        except Exception:  # noqa: BLE001  # silent-ok: a read-only handle close cannot affect the decoded result
            pass


def _scan_helper_tokens(path, side, wanted):
    """``{token: row}`` for the wanted tokens on one side's data sheet."""
    workbook = _open(path)
    try:
        if side not in workbook.sheetnames:
            return {}
        sheet = workbook[side]
        stream = sheet.iter_rows(values_only=True)
        header = list(next(stream, ()) or ())
        try:
            column = _unique_column(header, _KEY_HELPER_HEADER)
        except PublishedComparisonError:
            log.debug("published: %s has no unique %r column", side,
                      _KEY_HELPER_HEADER)
            return {}
        out = {}
        for row_number, raw in enumerate(stream, start=2):
            if raw is None or column >= len(raw):
                continue
            token = raw[column]
            if isinstance(token, str) and token in wanted:
                out[token] = row_number
                if len(out) == len(wanted):
                    break
        return out
    finally:
        try:
            workbook.close()
        except Exception:  # noqa: BLE001  # silent-ok: a read-only handle close cannot affect the decoded result
            pass
