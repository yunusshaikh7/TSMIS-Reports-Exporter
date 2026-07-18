"""Build the TSMIS-vs-TSN Ramp Summary discrepancy workbook (the AGGREGATE recipe).

Unlike the FLAT comparators (Ramp Detail etc.) that match per-postmile rows, the
two SUMMARY reports compare ONE statewide category-count table per side:

  * TSN side — the statewide "Ramp Summary Statewide" PDF: a single page of
    category counts (Highway Groups / On/Off / Population / Ramp Types) parsed with
    the same geometry helpers the per-route consolidator uses.
  * TSMIS side — the CONSOLIDATED Ramp Summary workbook: its per-route "TSAR Ramps
    Summary" sheet is SUMMED column-by-column into the same statewide totals (this
    is exactly what the workbook's live "Combined" sheet shows).

Each side reduces to {category-slug -> count}; the comparison key is the category
(has_route=False), the single data field is the count. The canonical category list
(the TSN 16-ramp-type superset, incl. the TSN-only P/V "Dummy" classes) lives in
summary_layout, shared with the familiar-layout sheet appended via extra_sheet_writer.

Reconciled on the 6.19 ground truth: TSN total 15410 vs TSMIS 15215; P (Dummy
Paired) and V (Dummy, Volume only) are TSN-only classifications, so they stay
one-sided ('Only in TSN' — CMP-AUD-024/025, never a fabricated TSMIS zero).
Both sides are independently validated before comparing (CMP-AUD-020/021/022):
strict whole-number counts, no duplicate categories, and the censused partition
contract (each block must reconcile against the grand total; the TSMIS
Ramp-Types block is bounded by the P/V-not-tabulated residual, which is exposed
as a note). Console-free; engine in compare_core.
"""
import re
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from dataclasses import replace

import compare_tsn_common as ctc
import consolidate_ramp_summary as rs
import consolidation_meta
import summary_layout
from compare_core import CompareSchema
from paths import today_str

REPORT_NAME = "Ramp Summary"
TSMIS_SHEET = rs.SUMMARY_SHEET_NAME          # "TSAR Ramps Summary" (per-route rows)
NORMALIZED_SHEET = "Ramp Summary (TSN)"      # the library's normalized 2-col workbook

_SPEC = summary_layout.RAMP_SUMMARY_SPEC
# (key, slug) for every compared category, in familiar display order.
_CATEGORIES = _SPEC.categories()

_SCHEMA = CompareSchema(
    report_name=REPORT_NAME,
    header=["Category", "Count"],
    side_a="TSMIS",
    side_b="TSN",
    id_noun="category",
    id_noun_plural="categories",
    sides_noun="systems",
    cmp_widths={"Count": 12},
    data_widths={"Count": 12},
    scope_flat="Statewide category totals",
    one_sided_note_extra=" (a category one system classifies and the other doesn't)",
    extra_sheet_writer=summary_layout.make_extra_sheet_writer(_SPEC),
)

# slug -> the consolidated workbook's row-2 display header (authoritative, from the
# consolidator's own GROUPS), so the TSMIS sum maps columns to the same slugs.
# COUNT columns only: the Source/Route/Audit columns carry text or formulas and
# must never be parsed as counts (CMP-AUD-021).
_COUNT_SLUGS = ({slug for _key, slug in _CATEGORIES}
                | {f.slug for f in _SPEC.footnotes})
_SLUG_TO_DISPLAY = {slug: disp for _grp, cols in rs.GROUPS for slug, disp in cols
                    if slug in _COUNT_SLUGS}

_NOLW = "ramp_points_no_linework"

# The censused partition contract (CMP-AUD-020), verified statewide on the 7.9
# ssor-prod consolidated set (126 routes) and the raw TSN statewide PDF: the
# TSMIS form counts its no-linework footnote inside the On/Off and Ramp-Type
# partitions but does not tabulate the TSN-only P/V dummy classes (their ramps
# are the bounded Ramp-Types residual — 22 on the censused pull, proved P=2/V=20
# by the same-pull Ramp Detail); the TSN statewide page partitions all four
# blocks exactly, P/V included.
_TSMIS_RULES = (
    summary_layout.SectionRule("Highway Groups", "exact"),
    summary_layout.SectionRule("On/Off Indicator", "exact", extra_slugs=(_NOLW,)),
    summary_layout.SectionRule("Population Groups", "exact"),
    summary_layout.SectionRule(
        "Ramp Types", "bounded", extra_slugs=(_NOLW,),
        reason="the TSMIS summary doesn't tabulate the TSN-only P/V dummy "
               "classes; those ramps appear in the 'Only in TSN' rows"),
)
_TSN_RULES = (
    summary_layout.SectionRule("Highway Groups", "exact"),
    summary_layout.SectionRule("On/Off Indicator", "exact"),
    summary_layout.SectionRule("Population Groups", "exact"),
    summary_layout.SectionRule("Ramp Types", "exact"),
)


# --------------------------------------------------------------------------- #
# TSN side: parse the statewide PDF -> {slug: count}
# --------------------------------------------------------------------------- #
def _tsn_data_page(pdf):
    """The page carrying the category table (the statewide PDF leads with cover /
    parameter pages); fall back to the last page."""
    for pg in pdf.pages:
        txt = pg.extract_text() or ""
        if re.search(r"Total number of Ramps", txt, re.I) or \
           ("Highway Groups" in txt and "Ramp Types" in txt):
            return pg
    return pdf.pages[-1]


def parse_tsn_pdf(path):
    """The statewide TSN Ramp Summary PDF -> {slug: count} over the full schema."""
    with pdfplumber.open(path) as pdf:
        page = _tsn_data_page(pdf)
        words = page.extract_words()
        left = rs.stitch_wrapped_rows(rs.get_rows_for_column(words, left=True))
        right = rs.stitch_wrapped_rows(rs.get_rows_for_column(words, left=False))
        rec, used = {}, set()
        rec.update(rs.match_schema(left, rs.HIGHWAY_GROUPS, used))
        rec.update(rs.match_schema(left, rs.ONOFF, used))
        rec.update(rs.match_schema(left, rs.POP_GROUPS, used))
        rec.update(rs.match_schema(right, rs.RAMP_TYPES))
        m = re.search(r"Total number of Ramps:\s*([\d,]+)", page.extract_text() or "", re.I)
        rec["total_ramps"] = int(m.group(1).replace(",", "")) if m else None
    return rec


def parse_tsn_source_claims(path):
    """The statewide TSN Ramp Summary print's source claims (CMP-AUD-146):
    the report identity/timing/submitter facts, required exactly-once across
    the document (the Ramp print has no fold or declared label correction, so
    identity is its whole claims record)."""
    path = Path(path)
    try:
        with pdfplumber.open(path) as pdf:
            full_text = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
    except Exception as e:
        raise ValueError(f"Could not read {path.name}: {type(e).__name__}: {e}")
    return {"schema_version": 1,
            "identity": ctc.tsn_print_identity(full_text, path.name)}


def claims_notes(claims, side_label="TSN"):
    """Human-readable exposure lines for the familiar sheet + log."""
    if not claims:
        return [f"{side_label} print: no source-claims record beside this "
                "normalized workbook (older normalization) — rebuild the TSN "
                "library to capture the print identity."]
    ident = claims.get("identity") or {}
    if not ident:
        return []
    return [f"{side_label} print identity: {ident.get('report_id')} · Event "
            f"{ident.get('event_id')} · reference {ident.get('reference_date')} "
            f"· submitted by {ident.get('submitter')} · generated "
            f"{ident.get('generated_time')} ({ident.get('location_criteria')})."]


def _load_tsn(path):
    """TSN side -> {slug: count}. Reads the raw statewide PDF, or the library's
    normalized 2-column workbook (Category | Count) if that was supplied."""
    path = Path(path)
    name = path.name
    if path.suffix.lower() == ".pdf":
        # The loader contract is ValueError-on-unreadable (run_files_compare
        # catches exactly that); a corrupt PDF must not escape as a raw
        # pdfplumber exception.
        try:
            return parse_tsn_pdf(path)
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"Could not read {name}: {type(e).__name__}: {e}")
    # normalized library workbook: a Category|Count sheet keyed on the compare key.
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        sn = NORMALIZED_SHEET if NORMALIZED_SHEET in wb.sheetnames else wb.sheetnames[0]
        key_to_slug = {k: s for k, s in _CATEGORIES}
        rec, seen = {}, set()
        it = wb[sn].iter_rows(values_only=True)
        next(it, None)                                   # header
        for r in it:
            if not r or r[0] is None:
                continue
            key = str(r[0]).strip()
            slug = key_to_slug.get(key)
            if slug is None:
                continue
            if key in seen:                              # CMP-AUD-022
                raise ValueError(f"{name} lists the category {key!r} twice — "
                                 "refusing an ambiguous normalized table")
            seen.add(key)
            v = r[1] if len(r) > 1 else None
            if v is None:
                raise ValueError(f"{name}: the category {key!r} has no count")
            rec[slug] = summary_layout.parse_count(v, source=name, category=key)
        return rec
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# TSMIS side: SUM the consolidated workbook's per-route sheet -> {slug: count}
# --------------------------------------------------------------------------- #
def _load_tsmis(path, note_sink=None):
    """TSMIS side -> {slug: count}. Sums each category column of the consolidated
    Ramp Summary workbook's per-route 'TSAR Ramps Summary' sheet (== its Combined
    sheet's live totals). Strict (CMP-AUD-021/022): count cells must be whole
    numbers (numeric text is parsed, never dropped; fractions and booleans
    refuse), a duplicated category column refuses, and a column the workbook
    lacks stays ABSENT — never a fabricated zero (reconcile_counts decides
    whether it was required). Blank cells (the never-tabulated P/V columns)
    contribute nothing. The route universe is validated before summing
    (CMP-AUD-071 — the mirror of Intersection Summary's CMP-AUD-183): a
    header-only or duplicate-route workbook refuses instead of certifying a
    phantom zero/double-counted universe, and the census status line lands in
    `note_sink`."""
    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        if TSMIS_SHEET not in wb.sheetnames:
            raise ValueError(f"{name} has no '{TSMIS_SHEET}' sheet — pick the "
                             "consolidated TSMIS Ramp Summary workbook.")
        it = wb[TSMIS_SHEET].iter_rows(values_only=True)
        next(it, None)                                   # row 1: group headers
        header = next(it, None) or ()                    # row 2: display headers
        names = [str(h).strip() if h is not None else None for h in header]
        if "Route" not in names:
            raise ValueError(f"{name} isn't a consolidated Ramp Summary workbook "
                             "(no 'Route' column) — consolidate the per-route "
                             "exports first.")
        route_col = names.index("Route")
        display_to_slug = {disp: slug for slug, disp in _SLUG_TO_DISPLAY.items()}
        col_slug, seen = {}, set()
        for i, h in enumerate(names):
            slug = display_to_slug.get(h) if h else None
            if slug is None:
                continue
            if slug in seen:                             # CMP-AUD-022
                raise ValueError(f"{name} has a duplicated category column "
                                 f"({h!r}) — refusing to sum an ambiguous table")
            seen.add(slug)
            col_slug[i] = (slug, h)
        sums = {slug: 0 for slug, _h in col_slug.values()}
        routes = []                                       # (row_no, Route cell)
        for row_no, row in enumerate(it, start=3):        # data starts at row 3
            if not row or all(c is None for c in row):
                continue
            routes.append((row_no, row[route_col] if route_col < len(row) else None))
            for i, (slug, disp) in col_slug.items():
                v = row[i] if i < len(row) else None
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue                             # blank cell (e.g. P/V)
                sums[slug] += summary_layout.parse_count(v, source=name,
                                                         category=disp)
        ctc.validate_route_universe(routes, name, path, note_sink)
        return sums
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# rows over the canonical category list
# --------------------------------------------------------------------------- #
def _rows(counts, side):
    """[[category_key, count], ...] for the categories EMITTED on `side`
    ('tsmis' | 'tsn'), in display order. Every emitted category must be PRESENT
    in `counts` (reconcile_counts guarantees it before this runs) — an absent
    category is a hard error, never a fabricated zero (CMP-AUD-020/021).
    Categories one system doesn't classify (P/V are TSN-only) are absent on the
    other side, so they land in 'Only in …' — matching the Intersection Summary
    recipe. Footnotes are NEVER emitted here (CMP-AUD-024): they are display-only
    and ride an out-of-band channel to the familiar sheet."""
    out = []
    for key, slug in _SPEC.categories_for(side):
        if slug not in counts:
            raise ValueError(f"the {side.upper()} table is missing the "
                             f"category {key!r} — cannot build comparison rows")
        out.append([key, counts[slug]])
    return out


def _footnote_values(tsmis_counts):
    """{footnote.key: value} for the display-only footnotes (e.g. Ramp Points w/out
    linework). CMP-AUD-024: passed out of band to the familiar sheet so a footnote can
    never become a one-sided comparison row or change the verdict."""
    out = {}
    for f in _SPEC.footnotes:
        v = tsmis_counts.get(f.slug)
        if v is not None:
            out[f.key] = int(v)
    return out


# --------------------------------------------------------------------------- #
# adapter surface (registry "files" kind)
# --------------------------------------------------------------------------- #
def suggest_name(tsmis_path):
    return f"TSMIS_vs_TSN_RampSummary_Comparison {today_str()}.xlsx"


def _load_pair(tsmis_path, tsn_path, footnote_sink=None, note_sink=None,
               events=None):
    """(rows_t, rows_n, warnings) for the shared driver. Each side emits only the
    categories it classifies (P/V are TSN-only, so they land in 'Only in TSN').

    Both inputs are INDEPENDENTLY validated before any row is built
    (CMP-AUD-020): every side-applicable category and the grand total must be
    present, and every block must partition the total per the censused
    SectionRule contract — a table that doesn't reconcile (e.g. all-zero
    categories under a non-zero total, a renamed header, dropped rows) refuses
    with a named block instead of comparing garbage. Bounded residuals (the
    TSMIS ramps in TSN-only P/V classes) are EXPOSED via `note_sink` onto the
    familiar sheet + the log — never fabricated into a category, and never a
    warning (warnings mean unreadable inputs). Display-only footnote values ride
    `footnote_sink` (CMP-AUD-024) — never the compared rows."""
    route_notes = []                                    # CMP-AUD-071 census line
    tsmis_counts = _load_tsmis(tsmis_path, note_sink=route_notes)
    tsn_counts = _load_tsn(tsn_path)
    tsmis_name, tsn_name = Path(tsmis_path).name, Path(tsn_path).name
    for slug, label in (("ramp_P_dummy_paired", "P - Dummy Paired"),
                        ("ramp_V_dummy_volume", "V - Dummy, Volume only")):
        if tsmis_counts.get(slug):
            raise ValueError(
                f"{tsmis_name} carries a non-zero '{label}' count "
                f"({tsmis_counts[slug]}) — the TSMIS summary is not supposed "
                "to tabulate the TSN-only P/V classes, so the one-sided "
                "comparison contract no longer holds; the comparator needs an "
                "update for this source change")
    # CMP-AUD-146: the print's identity claims — fresh from a raw PDF, or the
    # normalization sidecar's record (absent -> explicit diagnostic note).
    if Path(tsn_path).suffix.lower() == ".pdf":
        claims = parse_tsn_source_claims(tsn_path)
    else:
        claims = consolidation_meta.read_extra(tsn_path, "tsn_source_claims")
    notes = claims_notes(claims)
    notes += route_notes                                # CMP-AUD-071 route universe
    notes += summary_layout.reconcile_counts(
        _SPEC, tsmis_counts, "tsmis", _TSMIS_RULES,
        source=tsmis_name, side_label="TSMIS")
    notes += summary_layout.reconcile_counts(
        _SPEC, tsn_counts, "tsn", _TSN_RULES,
        source=tsn_name, side_label="TSN")
    if footnote_sink is not None:
        footnote_sink.clear()
        footnote_sink.update(_footnote_values(tsmis_counts))
    if note_sink is not None:
        note_sink.clear()
        note_sink.extend(notes)
    if events is not None:
        for n in notes:
            events.on_log(f"note: {n}")
    return _rows(tsmis_counts, "tsmis"), _rows(tsn_counts, "tsn"), []


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Ramp Summary TSMIS-vs-TSN AGGREGATE comparison workbook(s).
    `tsmis_path` is the consolidated TSMIS Ramp Summary workbook; `tsn_path` the
    TSN statewide PDF (or the library's normalized workbook)."""
    # CMP-AUD-024/020: per-run holders the loader fills and the familiar-sheet
    # writer reads, so footnotes + censused-residual notes reach the display
    # sheet without ever entering the compared universe.
    footnotes, notes = {}, []
    schema = replace(_SCHEMA, extra_sheet_writer=summary_layout.make_extra_sheet_writer(
        _SPEC, footnote_values=footnotes, extra_notes=notes))
    return ctc.run_files_compare(
        schema, tsmis_path, tsn_path, out_path,
        banner="Ramp Summary Comparison — TSMIS vs TSN (statewide category counts)",
        has_route=False,
        loader=lambda a, b: _load_pair(a, b, footnote_sink=footnotes,
                                       note_sink=notes, events=events),
        deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (pdfplumber, openpyxl).",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
