"""Build the TSMIS-vs-TSN Highway Log discrepancy workbook.

Takes a TSMIS Highway Log and a TSN Highway Log — either BOTH per-route
workbooks (31 columns, one route each) or BOTH consolidated workbooks (a
leading "Route" column, every route) — and writes the approved comparison
workbook: Summary / Spot Check / Comparison / (Routes) / Only in TSMIS /
Only in TSN / TSMIS / TSN, in the live-formulas and/or values flavors.

Since v0.10.0 the engine itself lives in compare_core.py (parameterized so
the cross-environment comparisons reuse it); this module is the Highway
Log's schema + input loading. The delegation is regression-verified: the
workbooks it writes are cell-for-cell identical to the pre-extraction
output (the format locked to the approved Route-1 sample). Comparison
semantics, sheet design and the formulas/values flavors are documented in
compare_core.py.

Console-free like the other report modules: progress via events.on_log,
overwrite via the confirm_overwrite callback, cancel honored between phases,
ConsolidateResult returned.
"""
import re
from dataclasses import replace
from pathlib import Path

try:
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import consolidation_meta
import consolidate_tsn_highway_log as tsn_hl   # marker/version + claims shape
import highway_log_columns as hlc       # the corrected column labels (one source)
from compare_core import CompareSchema, normalize_value
from compare_tsn_common import (make_notes_writer, row_has_data,
                                run_files_compare, suggest_route_name)

REPORT_NAME = "Highway Log"          # registry label (comparison type)
SHEET_NAME = "Highway Log"           # required sheet in both inputs

# The canonical (CORRECTED) per-route Highway Log layout. A workbook built before
# the label overhaul carries the old vendor labels — hlc.recognize() accepts it
# too and the engine compares by POSITION, relabeling to these for display.
# Consolidated workbooks carry ["Route"] + this.
EXPECTED_HEADER = hlc.HEADER

# The approved workbook's wording and geometry (see compare_core's CompareSchema):
# TSMIS/TSN side names, Med Wid zero-pad normalization (the corrected Median
# Width/Variance column), the Highway-Log-specific notes, the sample's widths,
# and the column tooltips + Legend sheet.
_SCHEMA = CompareSchema(
    report_name="Highway Log",
    header=EXPECTED_HEADER,
    side_a="TSMIS",
    side_b="TSN",
    id_noun="location",
    id_noun_plural="locations",
    pair_noun="postmile",
    sides_noun="systems",
    medwid_fields=(hlc.HEADER[19],),     # "Med Wid/Var [Med Wid]"
    date_fields=("Date of Rec", "Sig Chg. Date"),
    data_widths={"Location": 12, hlc.HEADER[1]: 11, "Description": 26, "Date of Rec": 11},
    cmp_widths={hlc.HEADER[1]: 12, "Description": 30, "Date of Rec": 12},
    one_sided_note_extra=" (mostly TSN segment splits and TSMIS realignment "
                         "markers)",
    trim_note_extra=" — the TSMIS export pads Description with trailing blanks",
    header_comment=hlc.comment_for,      # hover any column header for its meaning
    legend_writer=hlc.write_legend_sheet,  # a "Legend" tab explaining every column
    ditto_nonasserting=True,             # +/++/+++ = "see paired roadbed" -> never a diff
    ditto_resolver=hlc.display_fills,    # tint + hover the resolved value on each ditto cell
    key_normalizer=hlc.roadbed_canonical_location,  # unify roadbed encoding (TSMIS suffix vs TSN dittoed block)
)


def suggest_name(tsmis_path):
    """Output filename suggestion: 'TSMIS_vs_TSN_Route<id>_Comparison.xlsx'
    when the picked file carries a route token, consolidated-aware otherwise."""
    return suggest_route_name(tsmis_path, "Highway_Log", "TSMIS_vs_TSN")


_HL_WS_RE = re.compile(r"[\t\n\r\f\v]")


def _hl_normalize(v):
    """compare_core.normalize_value, plus: collapse tab/newline whitespace to a
    space. The TSMIS Excel export pads Description with trailing TAB characters,
    which Excel's TRIM (and _xl_trim) do NOT strip — so an otherwise-identical
    description ('END BR 5-95' vs 'END BR 5-95\\t\\t\\t') showed as a phantom
    difference. Replacing tabs with spaces at load lets TRIM collapse them, and
    keeps the values and formulas flavors in agreement (both then see only
    spaces). Highway-Log-scoped: other comparisons load through normalize_value
    directly and are unchanged."""
    nv = normalize_value(v)
    return _HL_WS_RE.sub(" ", nv) if isinstance(nv, str) else nv


def _load_input(path):
    """Load one Highway Log workbook -> (rows, has_route).

    Accepts the per-route layout (31 columns) and the consolidated layout
    ("Route" + 31). Raises ValueError with a user-safe message otherwise."""
    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"{name} has no '{SHEET_NAME}' sheet — pick a Highway Log "
                f"workbook (a TSMIS export or consolidation, or a TSN file "
                f"made by the TSN Highway Log consolidation).")
        rows_iter = wb[SHEET_NAME].iter_rows(values_only=True)
        header = [v for v in next(rows_iter, [])]
        while header and header[-1] in (None, ""):
            header.pop()
        # Accept the corrected labels OR the old vendor labels (a pre-overhaul
        # workbook) — the engine compares by POSITION and relabels to the
        # corrected header for display.
        has_route = hlc.recognize(header)
        if has_route is None:
            raise ValueError(
                f"{name} doesn't have the Highway Log column layout this "
                f"comparison expects — re-create it with this app, then retry.")
        n = len(header)
        rows = []
        for r in rows_iter:
            r = list(r)[:n] + [None] * max(0, n - len(r))
            if row_has_data(r):
                rows.append([_hl_normalize(v) for v in r])
        return rows, has_route
    finally:
        wb.close()


def _load_pair(path_a, path_b):
    """Load both Highway Log sides for the shared driver; both must have the
    SAME shape (per-route or consolidated). Also used by the two PDF-sourced
    flavors (compare_highway_log_pdf), which accept the identical layouts."""
    rows_a, route_a = _load_input(path_a)
    rows_b, route_b = _load_input(path_b)
    if route_a != route_b:
        per, con = ((path_b, path_a) if route_a else (path_a, path_b))
        raise ValueError(
            f"The two files have different shapes: {con.name} is a "
            f"consolidated workbook (has a Route column) but "
            f"{per.name} is per-route. Pick two per-route files or "
            f"two consolidated files.")
    return rows_a, rows_b, None, route_a


def _tsn_normalization_version(path):
    """The TSN-side workbook's declared normalization version (0 when the
    marker sheet is absent — the rows sheet kept its 31-column SHAPE across
    v5, so the marker is the only reliable signal on a bare file; the
    library path additionally auto-rebuilds via report_catalog, D2)."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # silent-ok: an unopenable file re-fails in _load_input with its own message
        return 0
    try:
        if tsn_hl.MARKER_SHEET not in wb.sheetnames:
            return 0
        for r in wb[tsn_hl.MARKER_SHEET].iter_rows(values_only=True):
            if r and str(r[0]).strip() == "Normalization version":
                try:
                    return int(r[1])
                except (TypeError, ValueError, IndexError):  # silent-ok: a malformed marker reads as version 0 — the caller then refuses with the rebuild hint (fail-safe)
                    return 0
        return 0
    finally:
        wb.close()


def require_current_tsn(path):
    """Refuse a pre-v5 TSN Highway Log workbook (CMP-AUD-157/045-HL): it
    merges detached suffixed-route sections (e.g. "07 LA 005 S" = route
    005S) into the base route, drops asterisk-leading printed Descriptions,
    and carries no conserved source claims — silently comparing it would
    resurrect the misattribution as false one-sided rows on both sides."""
    if _tsn_normalization_version(path) < tsn_hl.NORMALIZATION_VERSION:
        raise ValueError(
            f"{Path(path).name} was built by an older TSN Highway Log "
            "converter (pre-v5: suffixed-route sections like 005S merged "
            "into the base route, asterisk-leading descriptions dropped) — "
            "rebuild the TSN library and pick the fresh normalized workbook.")


def _load_pair_tsn(tsmis_path, tsn_path):
    """The vs-TSN pair loader: the TSN side must be a current (v5) normalized
    workbook; the TSMIS side is any recognized Highway Log export."""
    require_current_tsn(tsn_path)
    return _load_pair(tsmis_path, tsn_path)


# --------------------------------------------------------------------------- #
# Notes sheet — the conserved TSN source claims (CMP-AUD-157)
# --------------------------------------------------------------------------- #
_NOTES_TITLE = "Highway Log — TSMIS vs TSN: source claims"
_NOTES_LINES = (
    "Rows are keyed on Route + Location (roadbed-canonical). The TSMIS Highway "
    "Log export has no County column, so county can never be part of the "
    "two-sided key: the TSN print's district/county/route group ownership is "
    "conserved per document in the normalized workbook's sidecar (the numeric "
    "\"Cnty Odom\" column IS compared and participates in duplicate pairing).",
    "The print marks some sections with a detached route-suffix token "
    "(\"07 LA 005 S\") — those rows belong to the suffixed TSMIS route "
    "(005S), and the print's own COUNTY/ROUTE totals for them are all-zero.",
)


def claims_notes(claims, side_label="TSN"):
    """Human-readable exposure lines for the Notes sheet: the print identity
    the 12 districts agreed on, the suffixed-route sections, and the
    ADT/totals dispositions with their reconciliation summary."""
    if not claims:
        return [f"{side_label} print: no source-claims record beside this "
                "normalized workbook — rebuild the TSN library to capture "
                "the print identity, group ownership, ADT and totals claims."]
    docs = claims.get("documents") or []
    lines = [f"{side_label} print identity: {claims.get('report_id')} "
             f"{claims.get('report_title')} · report {claims.get('report_date')} "
             f"· cover year {claims.get('cover_year')} · "
             f"{len(docs)} district print(s)."]
    suffixed = claims.get("suffixed_route_sections") or []
    if suffixed:
        lines.append("Detached suffixed-route sections: "
                     + " · ".join(suffixed) + ".")
    tcu = sum((d.get("totals", {}).get("reconciliation", {})
               .get("tcu", {}).get("checked", 0)) for d in docs)
    r_chk = sum((d.get("totals", {}).get("reconciliation", {})
                 .get("route", {}).get("checked", 0)) for d in docs)
    r_ok = sum((d.get("totals", {}).get("reconciliation", {})
                .get("route", {}).get("exact", 0)) for d in docs)
    c_chk = sum((d.get("totals", {}).get("reconciliation", {})
                 .get("county", {}).get("checked", 0)) for d in docs)
    c_ok = sum((d.get("totals", {}).get("reconciliation", {})
                .get("county", {}).get("exact", 0)) for d in docs)
    lines.append(
        f"Printed totals: conserved by digest and typed; TOTAL = CONST + "
        f"UNCONST verified exactly on {tcu} mileage line(s); suffixed "
        f"sections verified all-zero. Route/county totals vs additive row "
        f"sums: {r_ok}/{r_chk} and {c_ok}/{c_chk} exact (measured — the "
        "print's odometer-based accounting is disclosed, not certified).")
    lines.append(
        "The print's three ADT columns (Look Back / P-S flag / Look Ahead) "
        "have no TSMIS counterpart: conserved by per-document digest in the "
        "sidecar, never compared, never fabricated.")
    return lines


def _schema_with_claims(tsn_path, schema=None):
    """The per-run schema: the Legend sheet plus a Notes sheet carrying the
    normalized workbook's persisted source claims (read from its sidecar —
    absent claims get an explicit rebuild hint instead of silence)."""
    base = schema if schema is not None else _SCHEMA
    claim_lines = claims_notes(
        consolidation_meta.read_extra(Path(tsn_path), "tsn_source_claims"))
    notes = make_notes_writer(_NOTES_TITLE,
                              tuple(_NOTES_LINES) + tuple(claim_lines))

    def _legend_and_notes(wb):
        hlc.write_legend_sheet(wb)
        notes(wb)

    return replace(base, legend_writer=_legend_and_notes)


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the comparison workbook(s). Returns a ConsolidateResult (same
    contract as the consolidators, so the GUI/console drive it identically).

    `mode`: "formulas" (the live workbook — every cell recalculates),
    "values" (same sheets and look, but the bulk is plain computed RESULTS —
    opens instantly, no F9), or "both" (two files: the picked name for the
    formulas copy and '<name> (values).xlsx' next to it)."""
    return run_files_compare(
        _schema_with_claims(tsn_path), tsmis_path, tsn_path, out_path,
        banner="Highway Log Comparison — TSMIS vs TSN",
        has_route=None, loader=_load_pair_tsn, deps_ok=_DEPS_OK,
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
