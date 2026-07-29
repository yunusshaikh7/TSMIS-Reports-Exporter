"""Build the ArcGIS-vs-TSN Clean Road CA HIGHWAYS discrepancy workbook.

Both sides carry the SAME 74-column `THY_*` shape: side A is OUR build from the
ArcGIS layer library (`consolidate_clean_highway` — role-marked with the
`ArcGIS Build` sheet), side B is the vendor's TSN `CA HIGHWAYS` extract (the
raw statewide `Sheet 1`, or the TSN library's normalized copy). One shared
projection loads either side; the role gates keep the sides honest (the ArcGIS
side REQUIRES the build marker, the TSN side REJECTS it), so an ArcGIS build
can never stand in for TSN or vice versa.

Row identity — the roadbed-aware physical span key (the Highway Detail canon
extended with the county, both sides carrying it natively):

    Route (name + suffix) · County · PM prefix · begin PM (decimal-canonical)
    · roadbed (the R/L/X PM suffix)

A row's END PM is deliberately NOT key material: where the two systems cut a
stretch differently, keying on the begin pairs the rows and surfaces the end
as a field difference instead of fabricating two one-sided rows.

Owner decisions (2026-07-22) carried here: every one of the 74 columns is
PRESENT; the columns with no TSMIS ArcGIS source (and TSN's own bookkeeping)
are CONTEXT — shown with TSN's value beside our empty cell, never counted as a
difference — and the Notes sheet indexes EVERY column back to its source layer
(the audit record). Value normalizations (dates to ISO, amounts to canonical
decimals, landmark edge trim) are format-only and documented in the Notes.

Console-free; engine in compare_core via compare_tsn_common.run_files_compare
(mode="both" writes the live-formulas workbook plus its values twin).
"""
import logging
import re
from dataclasses import replace
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import clean_highway_columns as chc
import compare_tsn_common as ctc
import comparison_contract as cc
from compare_core import CompareSchema
from consolidate_clean_highway import UNAVAILABLE_TOKEN
from paths import today_str

log = logging.getLogger(__name__)

REPORT_NAME = "Clean Road Highway"
ARC_SHEET = chc.ARC_SHEET
TSN_SHEET = chc.TSN_RAW_SHEET
NORMALIZED_SHEET = chc.NORMALIZED_SHEET
# CMP-AUD-037 discipline from day one: the library's normalized workbook is
# stamped with this version (tsn_load_clean_road mirrors it; the catalog's
# clean_highway normalization_version mirrors both).
NORMALIZATION_VERSION = 1
_NORMALIZED_SIDECARS = ()          # the normalized copy is the verbatim 74 cols

TSN_RAW_HEADER = tuple(chc.HEADER)
SHARED_HEADER = list(chc.HEADER)
KEY = "THY_BEGIN_PM_AMT"
KEY_FIELD = SHARED_HEADER.index(KEY)
CONTEXT_FIELDS = chc.CONTEXT_COLUMNS
DATE_FIELDS = ("THY_BEGIN_DATE", "THY_END_DATE", "THY_CREATE_DATE",
               "THY_LEFT_ROAD_EFF_DATE", "THY_MEDIAN_EFF_DATE",
               "THY_RIGHT_ROAD_EFF_DATE", "THY_ACCESS_EFF_DATE",
               "THY_LAST_SIG_CHG_DATE", "THY_RECORD_DATE", "THY_UPDATE_DATE",
               "THY_EXTRACT_DATE")
# Columns compared as canonical decimal amounts (float/int/text spellings of
# one number are the same value).
_AMOUNT_FIELDS = frozenset({
    "THY_BEGIN_OFFSET_AMT", "THY_END_OFFSET_AMT", "THY_SEG_ORDER_ID",
    "THY_END_PM_AMT", "THY_LENGTH_MILES_AMT", "THY_LT_LANES_AMT",
    "THY_LT_O_SHD_TOT_WIDTH_AMT", "THY_LT_O_SHD_TRT_WIDTH_AMT",
    "THY_LT_TRAV_WAY_WIDTH_AMT", "THY_LT_I_SHD_TOT_WIDTH_AMT",
    "THY_LT_I_SHD_TRT_WIDTH_AMT", "THY_MEDIAN_WIDTH_AMT",
    "THY_RT_LANES_AMT", "THY_RT_I_SHD_TOT_WIDTH_AMT",
    "THY_RT_I_SHD_TRT_WIDTH_AMT", "THY_RT_TRAV_WAY_WIDTH_AMT",
    "THY_RT_O_SHD_TOT_WIDTH_AMT", "THY_RT_O_SHD_TRT_WIDTH_AMT",
    "THY_DESIGN_SPEED_AMT", "THY_ADT_AMT", "THY_CHANGE_PER_MILE_AMT",
    "THY_TOLL_FOREST_CODE", "THY_CURB_LANDSCAPE_CODE",
    "THY_MAINT_SVC_LVL_CODE", "THY_NATIONAL_LANDS_CODE",
    "THY_SCENIC_FREEWAY_CODE",
})

def _provenance_table_lines():
    """The owner's column-match table (M2-E): per THY column — the SAME name on
    the TSN extract and our build (the build is THY-shaped to match) — the ArcGIS
    layer + source column it is PAINTED FROM, and whether it is COUNTED or shown as
    CONTEXT. So a reader sees exactly what matched what. Reads
    clean_highway_columns.PROVENANCE + CONTEXT_FIELDS; no count/pairing effect."""
    counted = set(chc.HEADER) - set(CONTEXT_FIELDS)
    out = []
    for name in chc.HEADER:
        _tier, layer, column, _note = chc.PROVENANCE[name]
        if layer and column:
            src = f"{layer} ({column})"
        elif layer:
            src = layer
        else:
            src = "no ArcGIS source"
        role = "counted" if name in counted else "context"
        out.append(f"    {name}  ·  built from {src}  ·  {role}")
    return tuple(out)


# HF-01: how many raw-source disagreements each surface itemizes inline. The
# Summary bullet is one sentence among many, so it stays short; the Notes
# sheet is the itemized record. Neither ever truncates silently — over the
# limit both say how many more there are and where to read them.
_SUMMARY_ITEM_LIMIT = 10
_NOTES_ITEM_LIMIT = 100

_NOTES_TITLE = "Clean Road Highway — ArcGIS build vs TSN: comparison notes"
_NOTES_LINES = (
        "Side A is OUR CA HIGHWAYS table, built from the owner's ArcGIS "
        "per-layer exports (arcgis_layers/) by the county+PM overlay "
        "consolidator; side B is the vendor's TSN CA HIGHWAYS extract. Both "
        "carry the same 74 THY_* columns.",
        "Rows are keyed on Route + County + PM prefix + begin postmile "
        "(decimal-canonical) + roadbed (the R/L/X PM suffix). The END "
        "postmile is deliberately not key material — where the two systems "
        "cut a stretch differently the rows still pair, and the end shows as "
        "a field difference instead of two one-sided rows.",
        "Dates are compared as ISO dates (format never counts); amount "
        "columns compare as canonical numbers ('02' = '2', '9.60' = '9.6'); "
        "the landmark text is edge-trimmed on both sides (the TSN extract "
        "pads with trailing blanks). THY_CHANGE_PER_MILE_AMT compares at 3 "
        "decimals: the extract's own per-row slope arithmetic wobbles in the "
        "4th decimal along one constant profile, and a real profile change "
        "moves it by whole units.",
        "The ADT profile family (Profile / ADT / Change-per-mile) IS "
        "compared and counted (owner decision 2026-07-22 — a wholesale "
        "column difference is exactly the signal this comparison exists to "
        "surface). Reading its differences, know the two model-fit classes "
        "inside the count alongside any real data changes: TSN's profiles "
        "CONTINUE ACROSS county lines (our build re-anchors at the county "
        "split), and where several count vintages are live TSN's choice is "
        "not always the latest year.",
        "THY_CITY_CODE is compared: the City layer carries city NAMES, and "
        "they are normalized to the TASAS letter codes via the built-in "
        "table (derived from statewide co-location, 99.92% agreement over "
        "21,906 voted rows); an unmapped name passes through verbatim so it "
        "surfaces as a visible difference instead of vanishing.",
        "CONTEXT columns (shown for reference, never counted as a "
        "difference): the TSN bookkeeping columns (id/element/lifecycle/"
        "create/update), the TASAS change-tracking flags, the columns with "
        "no TSMIS ArcGIS source (maintenance service level, the federal-aid "
        "trio, national lands, scenic freeway), THY_EXTRACT_DATE (ours is "
        "the build's as-of date by definition), and the two synthesized "
        "OFFSET columns (each side's offsets are its own derived cumulative; "
        "ours diverges from TSN's line wherever the two systems cut a "
        "stretch differently — that sliver already shows once, honestly, on "
        "END PM/LENGTH). Owner decision 2026-07-22: they stay PRESENT — "
        "both sides' values visible — so nothing is silently dropped.",
        "One-sided rows are stretches one side carries at a physical "
        "location (route + county + prefix + begin PM + roadbed) the other "
        "doesn't — segmentation differences show up here.",
        "Column match table — each THY column (the SAME name on the TSN extract "
        "and our build) · the ArcGIS layer (source column) it is painted from · "
        "whether it is COUNTED or shown as CONTEXT. The built workbook's "
        "Provenance sheet adds each layer's FeatureServer source + the per-column "
        "note:",
    ) + _provenance_table_lines()

_write_notes_sheet = ctc.make_notes_writer(_NOTES_TITLE, _NOTES_LINES)

_SCHEMA = CompareSchema(
    report_name=REPORT_NAME,
    header=SHARED_HEADER,
    side_a="ArcGIS",
    side_b="TSN",
    id_noun="segment",
    id_noun_plural="segments",
    pair_noun="postmile",
    sides_noun="systems",
    date_fields=DATE_FIELDS,
    data_widths={"THY_LANDMARK_SHORT_DESC": 26, "THY_BREAK_DESC": 10},
    cmp_widths={"THY_LANDMARK_SHORT_DESC": 30},
    one_sided_note_extra=(" (stretches one side carries at a physical "
                          "location the other doesn't)"),
    key_field=KEY_FIELD,
    context_fields=CONTEXT_FIELDS,
    # M2-E item 10 (owner, 2026-07-22): the full 74-column file is in the
    # sheet, so the 24 context (shown-only) column HEADERS are tinted grey
    # with a hover note — visibly distinct from the 50 counted columns, the
    # way one-sided ROWS are tinted. Presentation only; CRH-SW-E2 counts are
    # unchanged (re-proven on the real corpus 2026-07-23).
    context_header_fill="808080",
    legend_writer=_write_notes_sheet,
)

_ROUTE_RE = re.compile(r"^(\d+)([A-Z]?)$")


def _s(v):
    return "" if v is None else str(v).strip()


def _norm_route(name, suffix):
    s = _s(name)
    if s.endswith(".0"):
        s = s[:-2]
    m = _ROUTE_RE.match(s.upper())
    base = f"{int(m.group(1)):03d}{m.group(2)}" if m else s.upper()
    return base + _s(suffix).upper()


def _norm_date(v):
    """Both sides to YYYY-MM-DD: openpyxl datetime/date cells and the text
    forms iso_date already handles."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return ctc.iso_date(v)


def _norm_amount(v):
    """One canonical decimal text for a numeric cell in any spelling: 2 / 2.0
    / '02' / '2.40' -> '2' / '2' / '2' / '2.4'. Non-numeric values pass
    through stripped, so nothing is invented."""
    s = _s(v)
    if not s:
        return ""
    try:
        f = float(s)
    except ValueError:
        return s
    if f != f or f in (float("inf"), float("-inf")):
        return s
    text = f"{f:.10f}".rstrip("0").rstrip(".")
    if text in ("-0", ""):
        text = "0"
    return text


def _norm_cell(field, v):
    if field in DATE_FIELDS:
        return _norm_date(v)
    if field == "THY_CHANGE_PER_MILE_AMT":
        # The extract's own per-row slope arithmetic wobbles in the 4th
        # decimal (998.4636…998.4641 along ONE constant profile on route
        # 001); a real profile change moves this by whole units. Compared at
        # 3 decimals so the wobble never buries a real difference — the
        # Notes name the rule.
        s = _norm_amount(v)
        try:
            return _norm_amount(round(float(s), 3)) if s else s
        except ValueError:
            return s
    if field in _AMOUNT_FIELDS:
        return _norm_amount(v)
    return _s(v)


def _physical_span_key(route, county, prefix, begin_raw, roadbed, source_hint):
    """The typed physical identity of one clean-road span row (D4): route +
    county + prefix + decimal-canonical begin PM + roadbed. The displayed key
    text carries the same components so the Comparison sheet reads naturally."""
    if not county:
        raise ValueError(
            f"Clean Road Highway row (route {route}, PM {_s(begin_raw)}) has "
            f"no usable county in {source_hint} — cannot key it to a physical "
            "location")
    numeric = ctc.decimal_pm(begin_raw)
    component = f"{prefix}{numeric}{roadbed}"
    identity = cc.make_physical_identity(
        route, county, component,
        (cc.RawIdentityClaim("route", route),
         cc.RawIdentityClaim("county", county),
         cc.RawIdentityClaim("postmile_prefix", prefix),
         cc.RawIdentityClaim("postmile", _s(begin_raw)),
         cc.RawIdentityClaim("roadbed", roadbed)),
        f"{route} / {county} / {component}")
    return cc.physical_key(component, identity)


def _thy_row(vals, source_hint):
    """Project one 74-cell THY-shaped row (either side) onto
    [route, *SHARED_HEADER] with the begin-PM cell as the physical key."""
    h = {name: i for i, name in enumerate(chc.HEADER)}

    def g(name):
        i = h[name]
        return vals[i] if i < len(vals) else None

    route = _norm_route(g("THY_ROUTE_NAME"), g("THY_ROUTE_SUFFIX_CODE"))
    county = _s(g("THY_COUNTY_CODE")).upper()
    prefix = _s(g("THY_PM_PREFIX_CODE")).upper()
    roadbed = _s(g("THY_PM_SUFFIX_CODE")).upper()
    key = _physical_span_key(route, county, prefix, g("THY_BEGIN_PM_AMT"),
                             roadbed, source_hint)
    out = [route]
    for name in SHARED_HEADER:
        if name == KEY:
            out.append(key)
        else:
            out.append(_norm_cell(name, g(name)))
    return out


# --------------------------------------------------------------------------- #
# role gates (the CMP-AUD-066 pattern): the ArcGIS side must BE our build;
# the TSN side must NOT be.
# --------------------------------------------------------------------------- #
def _arc_marker_present(wb):
    return chc.ARC_MARKER_SHEET in wb.sheetnames


def _load_arc(path):
    """Side A: OUR ArcGIS-built workbook — the exact 74-column sheet plus the
    build marker (an unmarked THY-shaped workbook could be the TSN extract
    itself, and comparing TSN against TSN would certify a match no ArcGIS
    data ever entered)."""
    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        if not _arc_marker_present(wb):
            raise ValueError(
                f"{name} does not carry the '{chc.ARC_MARKER_SHEET}' marker, "
                "so it cannot stand as the ArcGIS side — build the Clean Road "
                "Highway workbook from the ArcGIS tab and pick that file.")
        if ARC_SHEET not in wb.sheetnames:
            raise ValueError(f"{name} has no '{ARC_SHEET}' sheet — rebuild "
                             "the Clean Road Highway workbook.")
        it = wb[ARC_SHEET].iter_rows(values_only=True)
        header = [_s(c) for c in (next(it, None) or ())]
        if header != chc.HEADER:
            raise ValueError(
                f"{name} does not carry the exact 74-column THY header — "
                "rebuild the Clean Road Highway workbook with this version.")
        return [_thy_row(list(r), f"{name} ({ARC_SHEET})")
                for r in it if ctc.row_has_data(r)]
    finally:
        wb.close()


def _load_tsn(path):
    """Side B: the TSN CA HIGHWAYS extract — the raw statewide `Sheet 1`, or
    the TSN library's normalized copy (marker-gated). A workbook carrying the
    ArcGIS build marker is refused — it is our side, not TSN's."""
    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        if _arc_marker_present(wb):
            raise ValueError(
                f"{name} is an ArcGIS-built Clean Road workbook (it carries "
                f"the '{chc.ARC_MARKER_SHEET}' marker), so it cannot stand as "
                "the TSN side — pick the TSN CA HIGHWAYS extract or the TSN "
                "library's normalized copy.")
        if NORMALIZED_SHEET in wb.sheetnames:
            it = wb[NORMALIZED_SHEET].iter_rows(values_only=True)
            header = [_s(c) for c in (next(it, None) or ())]
            ctc.require_shared_header_prefix(
                header, chc.HEADER, _NORMALIZED_SIDECARS, name, REPORT_NAME)
            ctc.require_current_normalization(
                wb, name, NORMALIZATION_VERSION,
                "pre-v1: no in-workbook normalization marker")
            return [_thy_row(list(r), f"{name} ({NORMALIZED_SHEET})")
                    for r in it if ctc.row_has_data(r)]
    finally:
        wb.close()
    return tsn_rows_from_raw(path)


def tsn_rows_from_raw(path):
    """Every row from the exact raw TSN statewide workbook."""
    with ctc.exact_raw_rows(
            path, TSN_SHEET, TSN_RAW_HEADER, REPORT_NAME,
            required_nonblank=("THY_COUNTY_CODE", "THY_ROUTE_NAME",
                               "THY_BEGIN_PM_AMT")) as (_header, rows_in):
        return [_thy_row(list(r), f"{Path(path).name} ({TSN_SHEET})")
                for r in rows_in]


# --------------------------------------------------------------------------- #
# adapter surface
# --------------------------------------------------------------------------- #
def suggest_name(_arc_path=None):
    return f"ArcGIS_vs_TSN_CleanRoadHighway_Comparison {today_str()}.xlsx"


def _load_pair(arc_path, tsn_path):
    rows_a = _load_arc(arc_path)
    rows_b = _load_tsn(tsn_path)
    return rows_a, rows_b, None


def _withheld_values(wb):
    """The build's itemized marked-anchor table as {(route, county, key
    component, column): [(withheld source value, the span's known postmile),
    ...]} — WHAT each unavailable marker stands in front of, keyed the way
    THIS comparison keys a row. Several unplaceable spans can anchor at one
    cell, so the value is a LIST: that cell's stretch genuinely carries more
    than one source value and none of them can be placed. Empty for a build
    that predates the table (its markers stay aggregate-only disclosure,
    exactly as before)."""
    if chc.ARC_MARKED_SHEET not in wb.sheetnames:
        return {}
    out = {}
    try:
        for r in wb[chc.ARC_MARKED_SHEET].iter_rows(min_row=2,
                                                    values_only=True):
            if not r or r[0] is None or len(r) < 7:
                continue
            route, county, prefix, begin_pm, roadbed, column, value = r[:7]
            station = r[8] if len(r) > 8 else None
            component = (_s(prefix).upper() + ctc.decimal_pm(begin_pm)
                         + _s(roadbed).upper())
            out.setdefault((_s(route).upper(), _s(county).upper(), component,
                            _s(column)), []).append((value, _s(station)))
    except Exception as e:      # silent-ok: an unreadable/malformed detail table costs the ITEMIZED disclosure only — the aggregate counts and the non-asserting rule still stand
        log.info("clean-road marked-anchor detail unreadable (%s: %s)",
                 type(e).__name__, str(e).splitlines()[0] if str(e) else "")
        return {}
    return out


def _build_skip_facts(path):
    """The built workbook's own skipped-source record from its marker sheet
    (HF-01): (skipped span count, marked anchor cell count, the withheld-value
    map). Zeroes and an empty map when the marker predates the record — an
    older build carries no tokens, so the plain schema is exactly right for
    it — or when nothing was skipped."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:      # silent-ok: the loader's role gate re-opens the file and surfaces the real failure with the actionable message
        log.info("clean-road skip facts unreadable (%s: %s)",
                 type(e).__name__, str(e).splitlines()[0] if str(e) else "")
        return 0, 0, {}
    try:
        if chc.ARC_MARKER_SHEET not in wb.sheetnames:
            return 0, 0, {}
        facts = {}
        for r in wb[chc.ARC_MARKER_SHEET].iter_rows(values_only=True):
            if r and r[0] is not None:
                facts[str(r[0]).strip()] = r[1] if len(r) > 1 else None

        def _count(key):
            try:
                n = int(facts.get(key))
            except (TypeError, ValueError):  # silent-ok: an absent/malformed marker count means an older or skip-free build — the plain schema is exactly right for it
                return 0
            return max(n, 0)

        return (_count("Skipped source spans"), _count("Marked anchor cells"),
                _withheld_values(wb))
    finally:
        wb.close()


def _pm_order(pm):
    """Sort anchors along the road, text-last so an unparseable one is stable."""
    try:
        return (0, float(pm), "")
    except (TypeError, ValueError):
        return (1, 0.0, str(pm))


def _conflict_text(item):
    """One raw-source disagreement as its disclosure sentence — every value
    the marker stands in front of, at the postmile the source DID give."""
    return (f"{item['route']} / {item['county']} / {item['location']} · "
            f"{item['field']}: ArcGIS source {item['source']}, TSN "
            f"{item['tsn']}")


def _source_conflicts(rows_a, rows_b, withheld):
    """Classify every unavailable marker against the TSN row this comparison
    pairs it with (HF-01 criterion 7). A marker asserts nothing either way,
    but a reader must still be told WHICH markers stand in front of a source
    value the TSN extract does not show — those are real facts a bare marker
    would otherwise hide. A cell agrees only when EVERY value withheld there
    is the value TSN already shows; anything else is itemized, including the
    stretch that genuinely carries two unplaceable values. Returns the
    itemized facts plus the exact census of every other class, so nothing is
    dropped silently."""
    key_at = 1 + KEY_FIELD
    county_at = 1 + SHARED_HEADER.index("THY_COUNTY_CODE")
    by_key = {}
    for r in rows_b:
        by_key.setdefault((r[0], r[key_at]), []).append(r)
    items = []
    census = {"agreed": 0, "unpaired": 0, "unrecorded": 0, "duplicated": 0}
    for r in rows_a:
        if UNAVAILABLE_TOKEN not in r:
            continue
        peers = by_key.get((r[0], r[key_at])) or ()
        route, county = _s(r[0]).upper(), _s(r[county_at]).upper()
        location = str(r[key_at])
        for i, v in enumerate(r):
            if i == 0 or v != UNAVAILABLE_TOKEN:
                continue
            field = SHARED_HEADER[i - 1]
            vals = withheld.get((route, county, location, field))
            if not vals:
                census["unrecorded"] += 1
                continue
            if not peers:
                census["unpaired"] += 1       # one-sided: no TSN row to differ from
                continue
            if len(peers) > 1:
                census["duplicated"] += 1     # duplicate keys: no single counterpart
                continue
            theirs = peers[0][i]
            anchors = sorted(((_norm_cell(field, one), pm)
                              for one, pm in vals),
                             key=lambda a: (_pm_order(a[1]), a[0]))
            if {one for one, _pm in anchors} == {theirs}:
                census["agreed"] += 1
                continue
            items.append(
                {"route": r[0], "county": county, "location": location,
                 "field": field, "tsn": theirs,
                 "source": ", ".join(f"{one} @ {pm}" if pm else str(one)
                                     for one, pm in anchors)})
    items.sort(key=lambda it: (it["route"], it["county"],
                               _pm_order(it["location"]), it["location"],
                               it["field"]))
    return dict(census, items=items, differs=len(items))


def _anomaly_text(census):
    """The classes that cannot be classified against a single TSN cell, named
    only when they exist — a duplicate key with no single counterpart, or a
    marker the build recorded no source value for."""
    parts = [(census["duplicated"], "sit on a key the TSN extract lists more "
                                    "than once"),
             (census["unrecorded"], "have no matching entry in the build's "
                                    f"'{chc.ARC_MARKED_SHEET}' record")]
    named = [f"{n} {why}" for n, why in parts if n]
    return (" " + "; ".join(named) + ", so they are left unclassified."
            if named else "")


def _coverage_sentence(skipped, marked, census):
    """The shared skipped-source facts: how many spans were unplaceable, how
    many anchors that marked, and how the marked anchors classify against
    TSN."""
    head = (f"the ArcGIS build could not place {skipped} source span(s) whose "
            f"begin or end postmile is unreadable (LocError=NO ERROR rows "
            f"with a missing PM endpoint); {marked} anchor cell(s) show "
            f"'{UNAVAILABLE_TOKEN}' and are never asserted or counted as "
            f"differences.")
    if census is None:
        return head
    compared = census["agreed"] + census["differs"]
    return (head + f" Of them {compared} are paired with a TSN row: "
            f"{census['agreed']} withhold only the value TSN already shows, "
            f"and {census['differs']} withhold at least one value TSN does "
            f"not."
            + (f" {census['unpaired']} have no TSN counterpart at all."
               if census["unpaired"] else "")
            + _anomaly_text(census))


def _itemized_text(items, limit):
    """The disagreements as one sentence, capped — and the cap is always
    stated, never a silent truncation."""
    shown = "; ".join(_conflict_text(it) for it in items[:limit])
    if len(items) <= limit:
        return shown
    return (f"{shown}; …and {len(items) - limit} more (the Notes sheet lists "
            "every one)")


def _summary_note(skipped, marked, census):
    """The Summary-sheet disclosure line. Resolved at write time, because the
    itemized disagreements are only knowable once BOTH sides are loaded — the
    inputs are digested before the loader reads them, so they cannot be read
    early just to compose this sentence."""
    note = "SOURCE COVERAGE — " + _coverage_sentence(skipped, marked, census)
    if census and census["items"]:
        note += (" Those source facts, itemized because a marker would "
                 "otherwise hide them: "
                 + _itemized_text(census["items"], _SUMMARY_ITEM_LIMIT) + ".")
    return (note + f" The built workbook's '{chc.ARC_MARKER_SHEET}' sheet "
            f"lists every skipped span"
            + (f" and '{chc.ARC_MARKED_SHEET}' every marked anchor"
               if census else "")
            + "; the Notes sheet explains the rule.")


def _disclosure_lines(skipped, marked, census):
    """The Notes-sheet disclosure block (HF-01): the skipped-source facts, the
    non-asserting rule, and every raw-source disagreement itemized, ahead of
    the standard notes so none of it can be buried under the 74-line column
    table."""
    lines = [
        "⚠ SOURCE COVERAGE — " + _coverage_sentence(skipped, marked, census)
        + " The raw rows carry usable AR/odometer measures, but the "
        "county+postmile contract never guesses a position from those "
        "calibrations, so each span's values are withheld at its one known "
        "endpoint instead.",
        f"Cells showing '{UNAVAILABLE_TOKEN}' are NON-ASSERTING (state N): "
        "the ArcGIS side HAS a source row there whose exact placement is "
        "unknowable, so the cell is neither '(blank)' (empty in the system) "
        "nor a countable difference — it is excluded from every difference "
        "count. The built workbook's '" + chc.ARC_MARKER_SHEET + "' sheet "
        "lists every skipped span with the measures it did have"
        + (f", and its '{chc.ARC_MARKED_SHEET}' sheet every marked anchor "
           "with the value withheld there." if census else "."),
    ]
    items = (census or {}).get("items") or ()
    if items:
        lines.append(
            f"The {len(items)} marked anchor(s) that withhold a value TSN "
            "does not show are named below, each with the postmile the "
            "source DID give. They stay non-asserting — an unplaceable span "
            "cannot be asserted at a position it never gave — but they are "
            "real source facts, so they are stated instead of hidden behind "
            "the marker:")
        lines.extend(f"     – {_conflict_text(it)}"
                     for it in items[:_NOTES_ITEM_LIMIT])
        if len(items) > _NOTES_ITEM_LIMIT:
            lines.append(f"     – …and {len(items) - _NOTES_ITEM_LIMIT} more "
                         f"(the build's '{chc.ARC_MARKED_SHEET}' sheet lists "
                         "every marked anchor).")
    return tuple(lines)


def _schema_for(arc_path, facts):
    """The per-run schema: when the built workbook records skipped source
    spans, opt into the non-asserting unavailable token and carry the exact
    counts — and the itemized raw-source disagreements — into the Summary note
    and the Notes sheet. `facts` is the box the loader fills once both sides
    are loaded; the Summary note and Notes sheet read it at write time. An
    older or skip-free build gets the plain module schema — byte-identical
    behavior."""
    skipped, marked, withheld = _build_skip_facts(arc_path)
    if not skipped and not marked:
        return _SCHEMA
    facts["withheld"] = withheld

    def notes_writer(wb):
        ctc.make_notes_writer(
            _NOTES_TITLE,
            _disclosure_lines(skipped, marked, facts.get("census"))
            + _NOTES_LINES)(wb)

    return replace(
        _SCHEMA,
        unavailable_rule=(UNAVAILABLE_TOKEN,
                          lambda: _summary_note(skipped, marked,
                                                facts.get("census"))),
        legend_writer=notes_writer)


def compare(arc_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Clean Road Highway ArcGIS-vs-TSN comparison workbook(s).
    `arc_path` is the ArcGIS-built workbook; `tsn_path` the TSN extract (raw
    or normalized). Returns a ConsolidateResult."""
    facts = {}

    def loader(a_path, b_path):
        rows_a, rows_b, warnings = _load_pair(a_path, b_path)
        # HF-01: classify the markers HERE — after the substrate has digested
        # both inputs, and off the rows it is about to compare, so the
        # disclosure describes exactly the data in the workbook.
        if facts.get("withheld"):
            facts["census"] = _source_conflicts(rows_a, rows_b,
                                                facts["withheld"])
        return rows_a, rows_b, warnings

    return ctc.run_files_compare(
        _schema_for(arc_path, facts), arc_path, tsn_path, out_path,
        banner="Clean Road Highway Comparison — ArcGIS build vs TSN",
        has_route=True, loader=loader, deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (openpyxl).",
        side_a="ArcGIS", side_b="TSN",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
