"""Build the TSMIS-vs-TSN Ramp Detail discrepancy workbook.

The reference v0.17.0 vs-TSN comparator (the recipe the other reports follow).
Both sides are XLSX, but in DIFFERENT shapes, so each side has its own loader that
projects to ONE shared comparison header keyed on PM (postmile):

  * TSMIS side — the CONSOLIDATED Ramp Detail workbook (sheet "TSAR - Ramp Detail"
    with a prepended "Route" column). Its header row has blank/merged cells that
    shift the City Code / R/U / Description LABELS right of their values, so the
    columns are read BY POSITION (verified against the real export), not by name.
  * TSN side — the statewide raw DB dump (sheet "Sheet 1", 18 named columns) OR the
    library's normalized workbook ("Route" + the shared header). Route comes from
    LOCATION ("01-DN-101" -> "101").

Reconciled on real data (route 1, 272/272 ramps match by normalized PM): PR, PM,
Date of Record (-> ISO), HG, Area 4, City Code, R/U (== TSN POP), and Description
(after stripping the TSMIS leading "<route>/" prefix) all line up; every residual
is a genuine TSMIS-vs-TSN difference. The TSN-only DB columns (Ramp Name, On/Off,
Ramp Type, ADT) have no TSMIS counterpart, so they are CONTEXT fields — shown for
reference, never counted as a diff (CompareSchema.context_fields).

Console-free like the other comparators: progress via events.on_log, overwrite via
the confirm_overwrite callback, a ConsolidateResult returned. Engine in compare_core.
"""
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import compare_tsn_common as ctc
import comparison_contract as cc
from compare_tsn_common import load_consolidated_rows, suggest_route_name
from compare_core import CompareSchema, normalize_value

REPORT_NAME = "Ramp Detail"
TSMIS_SHEET = "TSAR - Ramp Detail"       # the consolidated/per-route TSMIS sheet
TSN_SHEET = "Sheet 1"                     # the raw TSN statewide sheet

# The shared comparison header (key + fields), in display order. PM is the key.
# Compared fields assert; CONTEXT_FIELDS are shown but never counted as a diff.
# District is COMPARED (CMP-AUD-185): every source exposes it (inside Location /
# the library sidecar), and omitting it hid the one real District disagreement
# (005/SD/72.366 — TSMIS 12 vs TSN 11) behind an "identical" row.
KEY = "PM"
SHARED_HEADER = ["PR", "PM", "District", "Date of Record", "HG", "Area 4",
                 "City Code", "R/U", "Description", "Ramp Name", "On/Off",
                 "Ramp Type", "ADT"]
KEY_FIELD = SHARED_HEADER.index(KEY)      # 1
CONTEXT_FIELDS = ("Ramp Name", "On/Off", "Ramp Type", "ADT")   # TSN-only -> non-asserting
DATE_FIELDS = ("Date of Record",)
NORMALIZED_SHEET = "Ramp Detail (TSN)"    # the library's normalized TSN workbook sheet
TSN_RAW_HEADER = (
    "RAM_CONNECTION_ID", "RAMP_NANE", "LOCATION", "PR", "PM", "PM_SFX",
    "DATE_OF_RECORD", "HG", "AREA_4", "CITY_CODE", "POP", "ON_OFF",
    "ADT_EFF_YEAR", "ADT", "RAMP_TYPE", "EFF_DATE", "DESCRIPTION",
    "SEG_ORDER_ID",
)

# Notes sheet — the user-facing INDICATOR for the key/normalization/context
# choices (the same make_notes_writer legend the other vs-TSN comparators carry).
_write_notes_sheet = ctc.make_notes_writer(
    "Ramp Detail — TSMIS vs TSN: comparison notes",
    (
        "Rows are keyed on Route + County + PM (postmile, normalized to the TSN "
        "zero-padded form — '9.6' and '009.600' are the same ramp). County comes "
        "from each source's Location ('01-DN-101'); the same route+postmile in "
        "two counties is two different physical ramps, never paired.",
        "District IS compared (both sources carry it inside Location); the PM "
        "prefix (PR) and suffix are conserved source facts shown per side, not "
        "extra key components.",
        "Date of Record is compared as an ISO date — display-format differences "
        "(6/1/2024 vs 2024-06-01) never count.",
        "Description: the TSMIS export prepends the row's own route (\"001/NB "
        "OFF …\"); exactly that added prefix is stripped before comparing. A "
        "leading number TSN itself prints is authoritative data, never removed; "
        "only text EDGES are trimmed on both sides (two statewide TSN rows "
        "carry trailing tabs no representation prints).",
        "CONTEXT columns (shown for reference, never counted as a difference): "
        "Ramp Name, On/Off, Ramp Type and ADT are TSN database columns with no "
        "TSMIS counterpart — counting them would flood the workbook with one-sided "
        "cells that say nothing about agreement.",
        "One-sided rows are ramps one system lists at a physical location "
        "(route + county + postmile) the other doesn't.",
    ))

_SCHEMA = CompareSchema(
    report_name="Ramp Detail",
    header=SHARED_HEADER,
    side_a="TSMIS",
    side_b="TSN",
    id_noun="ramp",
    id_noun_plural="ramps",
    pair_noun="postmile",
    sides_noun="systems",
    date_fields=DATE_FIELDS,
    data_widths={"Description": 26, "Date of Record": 11},
    cmp_widths={"Description": 30, "Date of Record": 12},
    one_sided_note_extra=" (ramps one system lists at a postmile the other doesn't)",
    key_field=KEY_FIELD,
    context_fields=CONTEXT_FIELDS,
    legend_writer=_write_notes_sheet,
)

_ROUTE_FROM_LOCATION = re.compile(r"^\s*\d{2}-[A-Za-z]{2,3}-(\w+)\s*$")  # "01-DN-101" -> "101"
_DESC_PREFIX = re.compile(r"^\s*(\d+)\s*/\s*")    # TSMIS "001/NB OFF…" -> "NB OFF…"


def _dist_cnty(loc):
    """LOCATION '01-DN-101' / '04-CC.-004' -> ('01', 'DN'/'CC')."""
    parts = ("" if loc is None else str(loc)).strip().upper().split("-")
    dist = parts[0].strip() if parts else ""
    cnty = parts[1].strip().rstrip(".") if len(parts) >= 2 else ""
    return dist, cnty


def _physical_pm_key(route, county, pm_raw, claims, source_hint):
    """The D4 PhysicalKey for one Ramp Detail row (CMP-AUD-045): canonical
    identity is exactly the owner-approved `(Route, County, norm_pm(PM))` tuple
    (RD-79 accepted oracle) — the PM prefix and suffix stay conserved RAW CLAIMS,
    never key components (on the bound corpus PR differs on zero paired rows,
    and TSN's 313 print suffixes have no TSMIS counterpart; gluing either would
    fabricate one-sided rows out of physically identical ramps). The key's str
    payload is the normalized PM (what the side sheets display); a row whose
    county can't be established refuses loudly."""
    pm = _norm_pm(pm_raw)
    if not county:
        raise ValueError(f"Ramp Detail row (route {route}, PM {pm}) has no "
                         f"usable county in {source_hint} — cannot key it to a "
                         "physical location")
    identity = cc.make_physical_identity(
        route, county, pm,
        tuple(cc.RawIdentityClaim(name, value) for name, value in claims),
        f"{route} / {county} / {pm}")
    return cc.physical_key(pm, identity)


# --------------------------------------------------------------------------- #
# normalization (shared by both loaders)
# --------------------------------------------------------------------------- #
def _norm_route(tok):
    t = ("" if tok is None else str(tok)).strip().upper()
    m = re.fullmatch(r"(\d+)([A-Z]?)", t)
    return f"{int(m.group(1)):03d}{m.group(2)}" if m else t


# PM + Date-of-Record canon shared with Intersection Detail, homed in
# compare_tsn_common (P5b/S04). The rd._norm_pm / rd._iso_date names are kept so the
# loaders, importers, and the golden canary still resolve them.
_norm_pm = ctc.norm_pm
_iso_date = ctc.iso_date


def _strip_desc_prefix(text, route=None):
    """CMP-AUD-135: remove ONLY the TSMIS-export-added outer '<route>/' prefix.
    With `route` given (the comparison loaders), the leading number must BE the
    row's own route — a DIFFERENT numeric prefix is source data and survives
    (the accepted RD-79 oracle's declared TSMIS reading contract). Without a
    route (legacy/verification callers) any single leading 'digits/' strips —
    corpus-equivalent, since every TSMIS export line leads with its own route.
    OOXML escapes decode BEFORE the edge trim (CMP-AUD-197): the four Cactus
    City cells end `…_x000d_\\n` in the Excel export — an encoded CR the raw
    TSN extract never carries — and decode-then-trim removes exactly that
    export artifact while a trailing literal token would otherwise survive."""
    t = ctc.decode_ooxml_escapes("" if text is None else str(text)).strip()
    m = _DESC_PREFIX.match(t)
    if not m:
        return t
    if route is not None:
        rm = re.match(r"\d+", str(route))
        if rm is None or int(m.group(1)) != int(rm.group(0)):
            return t
    return t[m.end():]


def _edge_text(v):
    """Comparison-side Description projection: text edges are trimmed (tabs
    included) like the accepted oracle's declared value reading — the raw TSN
    extract carries trailing tabs on two censused route-126 rows that no TSMIS
    representation prints; INTERNAL whitespace still compares per D2 (no
    folding). Conservation is untouched: the normalized library and the raw
    claims keep the source text byte-for-byte."""
    v = _v(v)
    return v.strip() if isinstance(v, str) else v


def _v(x):
    """normalize_value + the installed-Excel OOXML decode (CMP-AUD-197):
    `_xHHHH_` escapes are how the Excel export serializes control characters
    (both hex cases; `_x005F_xHHHH_` stays a literal token), so reading them
    decoded IS reading the cell the way installed Excel displays it. The raw
    TSN extract carries none, so the TSN side is byte-unchanged; interior
    decoded characters survive as real compared content (no folding)."""
    nv = normalize_value(x)
    if not isinstance(nv, str):
        return nv
    return ctc.decode_ooxml_escapes(nv)


# --------------------------------------------------------------------------- #
# TSN side: raw statewide "Sheet 1" OR the library's normalized workbook
# --------------------------------------------------------------------------- #
def _raw_text(v):
    return "" if v is None else str(v)


def _tsn_raw_row(r, h):
    """Project one raw TSN Sheet-1 row (h = {NAME: idx}) to [route, *SHARED_HEADER].
    The PM cell is the D4 PhysicalKey (CMP-AUD-045); District is a compared field
    (CMP-AUD-185); the DESCRIPTION is preserved byte-for-byte — TSN's own leading
    numeric prefixes are authoritative data (CMP-AUD-135), only the TSMIS side
    strips its export-added route prefix."""
    def g(name):
        i = h.get(name)
        return r[i] if i is not None and i < len(r) else None
    _loc_raw = g("LOCATION")
    loc = "" if _loc_raw is None else str(_loc_raw)
    m = _ROUTE_FROM_LOCATION.match(loc)
    route = _norm_route(m.group(1)) if m else loc.strip().upper()
    district, county = _dist_cnty(loc)
    key = _physical_pm_key(route, county, g("PM"), (
        ("route", route), ("location", loc),
        ("postmile_prefix", _raw_text(g("PR"))),
        ("postmile", _raw_text(g("PM"))),
        ("postmile_suffix", _raw_text(g("PM_SFX")))), f"LOCATION {loc!r}")
    return [route,
            _v(g("PR")), key, district, _iso_date(g("DATE_OF_RECORD")),
            _v(g("HG")), _v(g("AREA_4")), _v(g("CITY_CODE")), _v(g("POP")),
            _edge_text(g("DESCRIPTION")),
            _v(g("RAMP_NANE")), _v(g("ON_OFF")), _v(g("RAMP_TYPE")), _v(g("ADT"))]


def require_tsn_raw_header(header):
    ctc.require_exact_raw_header(header, TSN_RAW_HEADER, REPORT_NAME)


def tsn_rows_from_raw(path):
    """Every route's rows from the exact raw statewide workbook."""
    with ctc.exact_raw_rows(
            path, TSN_SHEET, TSN_RAW_HEADER, REPORT_NAME,
            required_nonblank=("LOCATION", "PM")) as (header, rows_in):
        h = {n: i for i, n in enumerate(header)}
        return [_tsn_raw_row(r, h) for r in rows_in]


def _normalized_row(r):
    """Re-project one row from the normalized TSN-library sheet (v4: ['Route'] +
    SHARED_HEADER + the District/County/PM-Suffix sidecars) onto the shared
    shape, RE-APPLYING the FORMAT normalizations (PM zero-pad, ISO date) so a
    STALE library can't feed raw values through `_v` and flag phantom format
    diffs. The D4 PhysicalKey is REBUILT from the row's route + County sidecar +
    PM (CMP-AUD-045 — the sidecars are no longer sliced away), with the sidecar
    facts conserved as raw claims. Description is NOT re-stripped: the v4
    library preserves TSN's text byte-for-byte (CMP-AUD-135), and semantic
    normalizer changes are owned by the D2 `normalization_version` rebuild."""
    width = len(SHARED_HEADER) + 1
    vals = list(r)[:width + 3]                    # + District/County/PM-Suffix sidecars
    vals += [None] * (width + 3 - len(vals))
    route = "" if vals[0] is None else str(vals[0])
    pm_i = 1 + SHARED_HEADER.index("PM")
    date_i = 1 + SHARED_HEADER.index("Date of Record")
    district, county, pm_sfx = (_raw_text(vals[width]).strip(),
                                _raw_text(vals[width + 1]).strip(),
                                _raw_text(vals[width + 2]))
    key = _physical_pm_key(route, county, vals[pm_i], (
        ("route", route), ("district", district), ("county", county),
        ("postmile_prefix", _raw_text(vals[1])),
        ("postmile", _raw_text(vals[pm_i])),
        ("postmile_suffix", pm_sfx)), "the library's TSN County sidecar")
    out = [_v(c) for c in vals[:width]]
    out[pm_i] = key
    out[date_i] = _iso_date(vals[date_i])
    desc_i = 1 + SHARED_HEADER.index("Description")
    out[desc_i] = _edge_text(vals[desc_i])
    return out


def _load_tsn(path):
    """TSN side -> (rows, has_route=True). Reads the raw statewide Sheet-1, or the
    library's already-normalized workbook (header 'Route' + SHARED_HEADER + the
    v4 sidecars). A pre-v4 library (no District column / no PM-Suffix sidecar)
    predates the county-aware physical identity and REFUSES with a rebuild hint —
    comparing through it would silently re-weaken the key (CMP-AUD-045)."""
    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        sheets = wb.sheetnames
        if NORMALIZED_SHEET in sheets:                       # normalized library copy
            it = wb[NORMALIZED_SHEET].iter_rows(values_only=True)
            header = [("" if c is None else str(c).strip())
                      for c in (next(it, None) or ())]
            if "District" not in header or "TSN PM Suffix" not in header:
                raise ValueError(
                    f"{name} is an older normalized TSN Ramp Detail library "
                    "(before the county-aware physical identity) — rebuild the "
                    "TSN library and retry.")
            rows = [_normalized_row(r)
                    for r in it if r and any(c not in (None, "") for c in r)]
            return rows, True
    finally:
        wb.close()
    return tsn_rows_from_raw(path), True


# --------------------------------------------------------------------------- #
# TSMIS side: the consolidated Ramp Detail workbook (columns read BY POSITION)
# --------------------------------------------------------------------------- #
# Consolidated value positions (Route prepended): Route0 Location1 PR2 PM3 Date4
# suffix5 HG6 Area4 7 City8 R/U9 Description10 (blank11). The header LABELS shift
# right of City Code/R/U/Description, so position — not name — is authoritative.
def _tsmis_row(r):
    def at(i):
        return r[i] if i < len(r) else None
    route = _norm_route(at(0))
    loc = _raw_text(at(1)).strip()
    district, county = _dist_cnty(loc)
    key = _physical_pm_key(route, county, at(3), (
        ("route", route), ("location", loc),
        ("postmile_prefix", _raw_text(at(2))),
        ("postmile", _raw_text(at(3))),
        ("postmile_suffix", _raw_text(at(5)))), f"Location {loc!r}")
    return [route,
            _v(at(2)), key, district, _iso_date(at(4)),
            _v(at(6)), _v(at(7)), _v(at(8)), _v(at(9)),
            _strip_desc_prefix(at(10), route),               # the export-added prefix
            "", "", "", ""]                                  # TSN-only context: blank here


def _load_tsmis(path):
    """TSMIS side -> (rows, has_route=True). The CONSOLIDATED Ramp Detail workbook
    (a leading 'Route' column). Columns are read by position (the export's header
    row is column-shifted); a header sanity-check guards against layout drift."""
    return load_consolidated_rows(
        path, TSMIS_SHEET,
        missing_sheet_hint="pick the consolidated TSMIS Ramp Detail workbook.",
        bad_header_msg="isn't a CONSOLIDATED Ramp Detail workbook "
                       "(expected a leading 'Route' column) — consolidate the "
                       "per-route exports first.",
        # The export's header labels are column-shifted (read BY POSITION above),
        # so the gate checks shape, not exact labels: PM early + the full width.
        header_ok=lambda h: "PM" in h[:5] and len(h) >= 11,
        row_transform=_tsmis_row)


# --------------------------------------------------------------------------- #
# adapter surface (registry "files" kind)
# --------------------------------------------------------------------------- #
def suggest_name(tsmis_path):
    return suggest_route_name(tsmis_path, "Ramp_Detail", "TSMIS_vs_TSN_RampDetail")


def _load_pair(tsmis_path, tsn_path):
    """(rows_t, rows_n, warnings) for the shared driver — the FLAT detail pair carries
    no input warnings, so run_compare uses its () default."""
    rows_t, _ = _load_tsmis(tsmis_path)
    rows_n, _ = _load_tsn(tsn_path)
    return rows_t, rows_n, None


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Ramp Detail TSMIS-vs-TSN comparison workbook(s). `tsmis_path` is the
    consolidated TSMIS Ramp Detail workbook; `tsn_path` the TSN statewide (raw or
    normalized) workbook. Returns a ConsolidateResult."""
    return ctc.run_files_compare(
        _SCHEMA, tsmis_path, tsn_path, out_path,
        banner="Ramp Detail Comparison — TSMIS vs TSN", has_route=True,
        loader=_load_pair, deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (openpyxl).",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
