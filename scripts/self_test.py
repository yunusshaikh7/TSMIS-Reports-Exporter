"""Shared comprehensive runtime self-test for the bundled app.

Exercises EVERY real code path the frozen app depends on -- launch the system /
bundled browser + page.pdf() + a download, pdfplumber text/word/table extraction
(exactly what the Ramp Summary consolidator uses), an openpyxl write/read
round-trip, the dynamically-imported matrix modules (the F6 trio), the GUI bridge
(build the js_api + initial state, confirm the bundled ui/ assets), and --
best-effort -- a hidden WebView2 window cycle through the real JS bridge.

ONE body, two callers, so the exact shipped exe and the dev venv prove the same
thing (DRY):
  * scripts/gui_main.py  `--self-test`  -- the EXACT windowed release exe runs
    this before any real window is created; this is the build.ps1 -SelfTest
    release gate, exercising the precise artifact that ships (not a different
    console build -- the prior gate built a separate `full_smoke` exe).
  * build/full_smoke.py                 -- the dev/venv tool (same body; the
    render-stack step proves PIL/pypdfium2 WORK — they ship since v0.21.0).

`run(emit)` returns 0 on success and RAISES on any MANDATORY failure (import /
asset / registry / browser / pdf / openpyxl / dynamic-module). Only the hidden
WebView WINDOW probe may skip -- an environment that can't show a window is
tolerated; the import/registry sub-checks never skip. `emit` is a line sink
(defaults to print); gui_main passes a windowed-safe sink because the release exe
has no console/stdout.

Console-free-core note: this is a DIAGNOSTIC DRIVER (only gui_main and the build
tool import it), never imported by an engine module, and it writes only through
the injected `emit` -- so the no-console-in-core convention is preserved. The
heavy third-party imports live INSIDE run(), so merely importing this module
(e.g. the packaging reachability check, or normal startup) stays cheap.
"""
import sys
import tempfile
import threading
import time
from pathlib import Path

# A small page in the Ramp Summary shape (a heading + a table) -- enough to drive
# page.pdf() and exercise the exact pdfplumber calls the consolidator makes.
HTML = """
<h1>Route 005 Ramp Summary</h1>
<table border=1>
  <tr><th>Ramp</th><th>Count</th></tr>
  <tr><td>NB On</td><td>1234</td></tr>
  <tr><td>SB Off</td><td>5678</td></tr>
</table>
"""

# Flat modules the app imports DYNAMICALLY (the matrix tab loads them lazily) that
# a frozen bundle MUST carry -- the F6 trio. Importing them here is the runtime
# half of build/check_app_modules.py's offline packaging contract: if the bundle
# ever drops one, the exact shipped exe fails its own gate with a precise error.
# v0.21.0 adds the lazily-imported visual-evidence modules to the contract
# (v0.22.0: the Intersection Detail adapter joins; v0.24.0: Highway Log;
# v0.25.0: Highway Sequence + its lazily-resolved PDF consolidator/comparisons;
# v0.26.0: the Ramp Detail trio).
_DYNAMIC_REPORT_MODULES = ("matrix", "day_matrix", "pdf_excel_matrix",
                           "report_library",
                           "visual_evidence", "evidence_highway_detail",
                           "evidence_intersection_detail",
                           "evidence_highway_log",
                           "evidence_highway_sequence",
                           "consolidate_tsmis_highway_sequence_pdf",
                           "compare_highway_sequence_pdf",
                           "evidence_ramp_detail",
                           "consolidate_tsmis_ramp_detail_pdf",
                           "compare_ramp_detail_pdf")


def run(emit=None):
    """Run the comprehensive self-test. Returns 0 on success; raises on any
    mandatory failure. `emit(line)` is the output sink (default print). The temp
    workspace is ALWAYS removed -- on success and on a mandatory-failure raise
    (so a failing CI run doesn't leave diagnostics behind)."""
    import shutil
    emit = emit or print
    tmp = Path(tempfile.mkdtemp())
    try:
        return _exercise(tmp, emit)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _exercise(tmp, emit):
    """The comprehensive exercise body. run() owns the temp dir + its cleanup."""
    emit("=" * 60)
    emit("TSMIS Exporter -- full bundle self-test")
    emit("=" * 60)

    import openpyxl
    import pdfplumber
    from playwright.sync_api import sync_playwright
    from common import launch_browser                  # drives system Edge/Chrome

    emit(f"frozen={getattr(sys, 'frozen', False)}  "
         f"openpyxl={openpyxl.__version__}  pdfplumber={pdfplumber.__version__}")

    # 1. Chromium: launch + render to PDF (Ramp Summary path) + a download.
    pdf_path = tmp / "page.pdf"
    with sync_playwright() as p:
        browser = launch_browser(p, headless=True)    # system Edge/Chrome (or bundled)
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()
        page.set_content(HTML)
        page.pdf(path=str(pdf_path), format="Letter", print_background=True)
        page.set_content('<a id="d" href="data:text/plain,hello" download="x.txt">d</a>')
        with page.expect_download() as dl:
            page.click("#d")
        dl.value.save_as(str(tmp / "x.txt"))
        browser.close()
    assert pdf_path.stat().st_size > 0, "page.pdf produced nothing"
    emit(f"chromium: PDF {pdf_path.stat().st_size} bytes, download ok")

    # 2. pdfplumber: exactly the calls consolidate_ramp_summary makes.
    with pdfplumber.open(str(pdf_path)) as pdf:
        text = pdf.pages[0].extract_text() or ""
        words = pdf.pages[0].extract_words()
        tables = pdf.pages[0].extract_tables()
    assert "Route 005" in text, f"extract_text failed: {text!r}"
    assert any(w.get("text") == "1234" for w in words), "extract_words failed"
    emit(f"pdfplumber: text={len(text)} chars, words={len(words)}, tables={len(tables)}")

    # 3. openpyxl: write + read round-trip (consolidator output path).
    xlsx = tmp / "wb.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Route", "Ramp", "Count"]); ws.append(["005", "NB On", 1234])
    wb.save(str(xlsx))
    wb2 = openpyxl.load_workbook(str(xlsx))
    assert wb2.active["C2"].value == 1234, "openpyxl round-trip failed"
    emit("openpyxl: write/read round-trip ok")

    # 4. The visual-evidence render stack (SHIPS since v0.21.0 -- the excludes
    #    proof this step used to be is inverted): rasterize a PDF page
    #    (pdfplumber.to_image -> pypdfium2), draw the highlight box (PIL), and
    #    embed the PNG in a workbook (openpyxl's image insert needs PIL). These
    #    are exactly the calls visual_evidence makes, so a bundle missing any
    #    piece fails ITS OWN gate here, not the user's first evidence run.
    from PIL import ImageDraw
    from openpyxl.drawing.image import Image as XLImage
    with pdfplumber.open(str(pdf_path)) as pdf:
        img = pdf.pages[0].to_image(resolution=72).original.convert("RGB")
    ImageDraw.Draw(img).rectangle([10, 10, 60, 30], outline=(220, 20, 20), width=3)
    png = tmp / "evidence.png"
    img.save(str(png))
    assert png.stat().st_size > 0, "PDF page render produced nothing"
    wbe = openpyxl.Workbook()
    wbe.active.add_image(XLImage(str(png)), "A1")
    ev_xlsx = tmp / "evidence.xlsx"
    wbe.save(str(ev_xlsx))
    assert ev_xlsx.stat().st_size > 0, "evidence-workbook image embed failed"
    emit(f"evidence render stack: page->PNG {png.stat().st_size} bytes, "
         "highlight + workbook embed ok")
    emit(f"cryptography loaded (required by pdfminer): {'cryptography' in sys.modules}")

    # 5. The F6 trio: prove the frozen bundle carries the dynamically-imported
    #    matrix modules (precise failure here beats a confusing gui_api ImportError).
    import importlib
    for m in _DYNAMIC_REPORT_MODULES:
        importlib.import_module(m)
    emit(f"dynamic report modules import: {', '.join(_DYNAMIC_REPORT_MODULES)} ok")

    # 6. GUI bridge: js_api + initial state + bundled ui/ assets MUST work (these
    #    catch a prune/exclude that broke pywebview/pythonnet or lost the ui/ assets).
    import webview
    import gui_api

    class _NoWorker:                 # no background browser probe / GitHub update
        def __init__(self, *a, **k): pass     # the gate must be deterministic and
        def start(self): pass                 # offline-safe

    gui_api.CheckWorker = _NoWorker
    gui_api.UpdateWorker = _NoWorker

    api = gui_api.GuiApi()
    state = api.get_initial_state()
    assert state["reports"] and state["routes"], "GUI initial state incomplete"
    ui_index = gui_api._ui_index_path()
    assert ui_index.exists(), f"UI assets missing: {ui_index}"
    emit(f"gui: bridge api ok ({len(state['reports'])} reports, "
         f"{len(state['routes'])} routes, ui={ui_index})")

    # 7. Hidden WebView window cycle -- the ONLY skippable sub-check. The
    #    import/asset/registry checks above already passed and never skip; an
    #    environment that simply can't show a window (a headless CI box) is
    #    tolerated, since that is a display capability, not a packaging defect.
    res = {}

    def _drive(w):
        try:
            deadline = time.time() + 30
            while time.time() < deadline:                # wait for app.js to boot
                if w.evaluate_js("typeof window.__tsmis !== 'undefined'"):
                    break
                time.sleep(0.25)
            res["state"] = w.evaluate_js("window.__tsmis.test_state()")
        except Exception as e:
            res["err"] = f"{type(e).__name__}: {e}"
        finally:
            w.destroy()

    window = webview.create_window("smoke", str(ui_index), js_api=api, hidden=True)
    window.events.loaded += lambda: threading.Thread(
        target=_drive, args=(window,), daemon=True).start()
    watchdog = threading.Timer(60, lambda: (res.setdefault("err", "watchdog timeout"),
                                            window.destroy()))
    watchdog.daemon = True
    watchdog.start()
    try:
        webview.start(gui="edgechromium")
    except Exception as e:
        emit(f"gui: window skipped, environment can't start WebView2 "
             f"({type(e).__name__}: {e})")
    else:
        watchdog.cancel()
        if not res.get("state"):
            raise AssertionError(f"gui window cycle failed: {res.get('err', 'no JS state')}")
        emit(f"gui: WebView window + JS bridge ok ({res['state']})")

    emit("")
    emit("SMOKE OK -- every app-required code path works.")
    return 0
