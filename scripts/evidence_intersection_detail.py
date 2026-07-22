"""Intersection Detail adapter for the visual-evidence generator (visual_evidence).

Supplies everything report-specific the engine needs to turn a vs-TSN diff into
a pair of highlighted PDF snippets:

  * the DIFF SOURCE — both comparison sides re-loaded through the comparator's
    OWN loaders/normalizations (compare_intersection_detail_tsn) and compared
    with compare_core's cell trim, so an "example" is exactly a cell the
    comparison flags, never a re-derivation that could drift;
  * the TSMIS locator — the same document-grid walk the Intersection Detail PDF
    consolidator uses, kept in LOCKSTEP with it (see _locate note), but keeping
    each record's page / y-extent / column geometry so a field maps to a cell
    rectangle;
  * the TSN locator — the STATEWIDE TASAS print (one file, every district;
    per-district files work too) parsed on its fixed monospace column template,
    with word positions kept so a field maps to a box even when the cell is
    blank. The whole print is indexed ONCE per file (cached on content identity);
    the engine's per-district locate calls are then dictionary lookups.
  * VERIFICATION projections — a candidate is only usable when the value parsed
    back out of each PDF, run through the comparator's own projections, equals
    the value the comparison compared (so an evidence image can never show
    something other than what was diffed; the known render skews — the print's
    truncated long descriptions, its collapsed multi-spaces — are skipped with
    a recorded reason instead).

Row identity matches the comparison: (route, PM), restricted to keys UNIQUE on
both sides of a route so a highlight is THE row (the comparison's occurrence
pairing handles duplicates — cross-county postmile repeats included; evidence
just avoids them). Console-free; pdfplumber/openpyxl gated by the engine.
"""
import logging
import re
from collections import defaultdict
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import artifact_store
import compare_intersection_detail_tsn as idt
import consolidate_tsmis_intersection_detail_pdf as idpdf
from compare_core import _xl_trim, compared_cell, published_key_text
from pdf_table_lib import cluster_by_top, median, require_document_route
from tsn_load_intersection_detail import SIDECAR_HEADER, tsn_rows_with_dcr  # noqa: F401

log = logging.getLogger("tsmis.evidence")

REPORT_LABEL = "Intersection Detail"
KEY_LABEL = idt.KEY
# Every compared column except the key itself. Route Suffix included — its TSN
# side is boxed on the LOCATION cell (where the print carries the suffix).
FIELDS = [f for f in idt.SHARED_HEADER if f != idt.KEY]

_KEY_I = 1 + idt.KEY_FIELD           # PM's index in a loader row ([route, *header])


def _S(v):
    return ("" if v is None else str(v)).strip()


# --------------------------------------------------------------------------- #
# diff source — both sides through the comparator's own loaders
# --------------------------------------------------------------------------- #
def load_sides(consolidated_path, tsn_path):
    """(tsmis_rows, tsn_rows, sidecar, note) — both sides in the comparator's
    shape ([route, *SHARED_HEADER]). `sidecar` maps (route, key) -> [(district,
    county), ...] for locating TSN rows in the statewide print; None when the
    TSN workbook carries no district info (an old normalized library), with
    `note` saying what to do about it."""
    tsmis_rows, _route_keyed = idt._load_tsmis(consolidated_path)
    tsn_rows, sidecar, note = _load_tsn_with_sidecar(tsn_path)
    return tsmis_rows, tsn_rows, sidecar, note


def _load_tsn_with_sidecar(path):
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if idt.NORMALIZED_SHEET in wb.sheetnames:
            it = wb[idt.NORMALIZED_SHEET].iter_rows(values_only=True)
            header = [_S(c) for c in (next(it, []) or [])]
            # CMP-AUD-045: the county-aware identity needs the v3 sidecars —
            # same rule as the comparison loader, surfaced as the rebuild note.
            if "TSN County" not in header:
                return [], None, ("the normalized TSN library predates the "
                                  "evidence columns — rebuild the TSN library "
                                  "(Settings) and run the comparison again")
            rows, sidecar = [], defaultdict(list)
            for r in it:
                if not r or all(c in (None, "") for c in r):
                    continue
                row = idt._normalized_row(r)   # re-projected like the comparison
                rows.append(row)
                dist = _S(r[1 + len(idt.SHARED_HEADER)])
                cnty = _S(r[2 + len(idt.SHARED_HEADER)]).rstrip(".")
                sidecar[(row[0], row[_KEY_I])].append((dist, cnty))
            return rows, sidecar, None
    finally:
        wb.close()
    rows, dcr = tsn_rows_with_dcr(path)          # a RAW statewide file
    sidecar = defaultdict(list)
    for row, (dist, cnty) in zip(rows, dcr):
        sidecar[(row[0], row[_KEY_I])].append((dist, cnty))
    return rows, sidecar, None


def enumerate_diffs(tsmis_rows, tsn_rows, sidecar):
    """{field: [example]} over (route, PM) keys UNIQUE per route on BOTH sides —
    each example a cell the comparison COUNTS. Equality is the engine's OWN
    compared_cell verdict (loader projections + the Excel TRIM), so inequality
    here == a red cell there (CMP-AUD-107 — the shared engine)."""
    sc = idt._SCHEMA
    a_route, b_route = defaultdict(list), defaultdict(list)
    # The A-side row's position rides along so the engine can address the exact
    # CONSOLIDATED-workbook cell an Excel-compared value came from (CMP-AUD-210).
    for i, r in enumerate(tsmis_rows):
        a_route[r[0]].append((i, r))
    for r in tsn_rows:
        b_route[r[0]].append(r)
    diffs = defaultdict(list)
    for route in sorted(set(a_route) & set(b_route)):
        a_ct, b_ct = defaultdict(int), defaultdict(int)
        for _i, r in a_route[route]:
            a_ct[r[_KEY_I]] += 1
        for r in b_route[route]:
            b_ct[r[_KEY_I]] += 1
        a_by = {r[_KEY_I]: (i, r) for i, r in a_route[route]
                if a_ct[r[_KEY_I]] == 1}
        b_by = {r[_KEY_I]: r for r in b_route[route] if b_ct[r[_KEY_I]] == 1}
        for key in set(a_by) & set(b_by):
            (ia, ra), rb = a_by[key], b_by[key]
            pub_key = published_key_text(sc, ra)
            for i, f in enumerate(idt.SHARED_HEADER):
                if f == idt.KEY:
                    continue
                cell = compared_cell(sc, i, ra, rb, 1)
                if cell.verdict is False:
                    dist, cnty = (sidecar.get((route, key)) or [("", "")])[0]
                    # The engine's locators key on the plain normalized-PM text
                    # (+ county for TSN); the ID-79 PhysicalKey's str payload IS
                    # that text (CMP-AUD-045).
                    diffs[f].append(dict(
                        route=route, key=str(key), field=f,
                        va=cell.display_a, vb=cell.display_b,
                        dist=dist, cnty=cnty, row_index=ia,
                        pub_key=pub_key, display=cell.display))
    return diffs


# --------------------------------------------------------------------------- #
# verification projections
# --------------------------------------------------------------------------- #
def project(field, raw):
    """A raw PDF cell value -> the compared form, via the comparator's own
    per-field projection + compare_core's cell trim (Route Suffix is
    Location-derived, handled by the callers)."""
    return _xl_trim(idt._project(field, raw))


# --------------------------------------------------------------------------- #
# TSMIS side — the per-route "Intersection Detail (PDF)" export
# --------------------------------------------------------------------------- #
def tsmis_pdf_path(pdf_dir, route):
    return Path(pdf_dir) / f"intersection_detail_route_{route}.pdf"


# shared field -> (physical line, grid window) in the TSMIS print. Line 1 uses
# the 21-column rowA grid, line 2 the 18-column rowB grid (its window 3 is the
# merged Description). Route Suffix boxes the Location cell — that's where the
# print carries the suffix ('11 IMP 008U'); the site leaves the 'S' column blank.
_TSMIS_CELL = {
    "PR": (1, 0), "Route Suffix": (1, 3), "District": (1, 3), "County": (1, 3),
    "Date of Record": (1, 4), "HG": (1, 5),
    "City Code": (1, 6), "R/U": (1, 7),
    "INT Type Eff-Date": (1, 8), "INT Type": (1, 9),
    "Control Type Eff-Date": (1, 10), "Control Type": (1, 11),
    "Lighting Eff-Date": (1, 12), "Lighting": (1, 13),
    "ML Eff-Date": (1, 14), "ML Mastarm": (1, 15), "ML Left Chan": (1, 16),
    "ML Right Chan": (1, 17), "ML Traffic Flow": (1, 18), "ML Num Lanes": (1, 19),
    "Description": (2, 3), "Main Line Length": (2, 4),
    "CS Eff-Date": (2, 5), "CS Mastarm": (2, 6), "CS Left Chan": (2, 7),
    "CS Right Chan": (2, 8), "CS Traffic Flow": (2, 9), "CS Num Lanes": (2, 10),
    "Int St Eff-Date": (2, 11),
    "Intrte Route": (2, 12), "Intrte PM Prefix": (2, 14),
    "Intrte Postmile": (2, 15), "Intrte PM Suffix": (2, 16),
    "Xing Line Lgth": (2, 17),
}
# shared field -> its 35-column output-row position (the consolidated positions
# minus the leading Route column).
_TSMIS_SRC = {f: p - 1 for f, p in idt._TSMIS_POS.items()}


def _doc_edges(pdf, n_cols):
    """Median (x0,x1) CELL edges per column from the document's shaded bands —
    the box for a BLANK cell, where no character can supply a bbox. (The
    consolidator's windows are midpoint-contiguous, wider than the drawn cell.)"""
    bands = []
    for page in pdf.pages:
        by_top = defaultdict(list)
        for r in page.rects:
            w = r["x1"] - r["x0"]
            h = r["bottom"] - r["top"]
            if (idpdf.CELL_MIN_W < w < page.width - 10
                    and idpdf.CELL_MIN_H < h < idpdf.CELL_MAX_H):
                by_top[round(r["top"])].append(r)
        for cells in by_top.values():
            if len(cells) == n_cols:
                bands.append(sorted(cells, key=lambda r: r["x0"]))
    if not bands:
        return None
    return list(zip([median([b[i]["x0"] for b in bands]) for i in range(n_cols)],
                    [median([b[i]["x1"] for b in bands]) for i in range(n_cols)]))


def _lmeta(chars, page_no, win, edges):
    return {"page": page_no, "win": win, "edges": edges, "chars": list(chars),
            "top": min(c["top"] for c in chars),
            "bottom": max(c["bottom"] for c in chars)}


def locate_tsmis(pdf_path, needed_keys):
    """{normalized_pm: [record]} for `needed_keys`; a record carries the parsed
    35-column row plus per-line meta (page, y-extent, windows, edges, chars).

    LOCKSTEP: this walk mirrors consolidate_tsmis_intersection_detail_pdf
    .parse_pdf step for step (the document grids from both band shapes, the
    padded-postmile rowA test, the old-layout line skip, the integer-column-1
    rowB pairing that steps over page furniture, the CMP-AUD-049 cover
    ROUTE-parameter capture before the geometry gate) — it only ADDS position
    capture. A behavior change there must land here too; check_visual_evidence
    pins the shared pieces so a drift fails the gate.

    CMP-AUD-049 (evidence half): raises pdf_table_lib.RouteIdentityError when
    the cover's own "ROUTE : NNN" parameter doesn't confirm the route the
    filename names (the per-record Location cells canNOT identify the
    document — an intersection with another route prints the OTHER route's
    mainline Location; see idpdf.COVER_ROUTE_RE)."""
    found = defaultdict(list)
    doc_routes = set()                 # the cover's own ROUTE parameter (049)
    fm = re.search(r"route_([0-9A-Za-z]+)\.pdf$", str(pdf_path))
    file_route = fm.group(1) if fm else None

    def require_identity():
        require_document_route(
            Path(pdf_path).name,
            idpdf._norm_route(file_route) if file_route else None,
            [idpdf._norm_route(t) for t in doc_routes],
            claim_desc="the cover's \"ROUTE : NNN\" parameter")

    pending = pending_meta = None
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[:idpdf._COVER_SCAN_PAGES]:
            for _top, chars in idpdf._cluster_lines(page):
                cm = idpdf.COVER_ROUTE_RE.match(
                    "".join(c["text"] for c in chars))
                if cm:
                    doc_routes.add(cm.group(1))
            if doc_routes:
                break
        win_a, win_b = idpdf._doc_windows(pdf)
        if win_a is None:
            require_identity()         # a grid-less document must still be
            return found               # the route it is asked to verify
        edges_a = _doc_edges(pdf, idpdf.N_COLS_A)
        edges_b = _doc_edges(pdf, idpdf.N_COLS_B)
        for page_no, page in enumerate(pdf.pages, 1):
            for _top, chars in idpdf._cluster_lines(page):
                vals = idpdf._assign_columns(chars, win_a)
                if idpdf._is_rowA(vals):
                    pending = vals
                    pending_meta = _lmeta(chars, page_no, win_a, edges_a)
                    continue
                if (idpdf.OLD_PM_RE.match(vals[1])
                        and idpdf.LOCATION_RE.search(vals[3] or "")):
                    continue                     # a pre-update line (see parse_pdf)
                if pending is not None:
                    vals_b = idpdf._assign_columns(chars, win_b)
                    if idpdf.INT_ROWB_RE.match(vals_b[1] or ""):
                        row = idpdf._make_row(pending, vals_b)
                        key = idt._norm_pm(row[1])
                        if key in needed_keys:
                            found[key].append(
                                {"row": row, "m1": pending_meta,
                                 "m2": _lmeta(chars, page_no, win_b, edges_b)})
                        pending = pending_meta = None
    require_identity()
    return found


def tsmis_value(rec, field):
    """The compared value this PDF record carries for `field` (verification)."""
    if field == "Route Suffix":
        return idt._split_route(rec["row"][3])[1]
    if field in ("District", "County"):
        district, county = idt._dist_cnty(rec["row"][3])
        return district if field == "District" else county
    return project(field, rec["row"][_TSMIS_SRC[field]])


def tsmis_box(rec, field):
    """(page_no, cell_box, record_yspan, table_xspan) for `field`'s cell.
    Rejects (returns None) a record whose two lines landed on different pages —
    a single-page strip can't show both, and the record box would be wrong."""
    if rec["m1"]["page"] != rec["m2"]["page"]:
        return None
    line, idx = _TSMIS_CELL[field]
    meta = rec["m1"] if line == 1 else rec["m2"]
    lo, hi = meta["win"][idx]
    hits = [c for c in meta["chars"] if lo <= (c["x0"] + c["x1"]) / 2 < hi]
    if hits:
        x0, x1 = min(c["x0"] for c in hits), max(c["x1"] for c in hits)
    elif meta["edges"]:
        x0, x1 = meta["edges"][idx]
    else:
        return None
    e1, e2 = rec["m1"]["edges"], rec["m2"]["edges"]
    xs = [e[0][0] for e in (e1, e2) if e] + [e[-1][1] for e in (e1, e2) if e]
    if not xs:
        return None
    return (meta["page"],
            (x0 - 2, meta["top"] - 2, x1 + 2, meta["bottom"] + 2),
            (rec["m1"]["top"], rec["m2"]["bottom"]),
            (min(xs) - 4, max(xs) + 4))


# --------------------------------------------------------------------------- #
# TSN side — the statewide TASAS print (fixed monospace column template)
# --------------------------------------------------------------------------- #
# The print is line-printer output (Courier data lines; Times header furniture)
# on ONE fixed template document-wide: every field starts at the same x on every
# page (verified statewide against the raw extract). LOCATION is one left-packed
# field ('12 ORA 001' / '07 LA 001' — a 2-char county shifts the route token),
# so it is ONE window split by tokens. Words are assigned to windows by MAX
# OVERLAP: a signature-flagged date ('Y98-08-28') starts a character early and
# leans into its neighbor, but its overlap keeps it in the date window.
_L1_WIN = [
    ("PP", 10, 24), ("PM", 24, 62), ("LOC", 64, 148),
    ("DATE_REC", 150, 200), ("HG", 202, 213), ("CITY", 213, 246),
    ("RU", 256, 278), ("EFF_DATE_INT", 279, 331), ("TY_INT", 332, 348),
    ("EFF_DATE_CT", 349, 394), ("TY_CT", 395, 412), ("EFF_DATE_LT", 413, 460),
    ("LT_TY", 473, 491), ("EFF_DATE_ML", 492, 541), ("MAIN_SM", 542, 555),
    ("MAIN_LC", 556, 567), ("MAIN_RC", 568, 578), ("MAIN_TF", 579, 588),
    ("MAIN_NL", 588, 599), ("X_CROSS_OVERRIDE", 599, 623),
    ("MAIN_EFF_DATE", 624, 673), ("MAIN_ADT", 674, 716),
]
_L2_WIN = [
    ("DESCRIPTION", 64, 251), ("MAIN_OVERRIDE", 251, 275),
    ("CROSS_BEGIN_DATE", 279, 331), ("CS_SM", 331, 345), ("CS_LC", 345, 357),
    ("CS_RC", 357, 369), ("CS_TF", 369, 381), ("CS_NL", 381, 395),
    ("EFF_DATE", 410, 458), ("CROSS_ADT", 458, 493),
    ("CR_ROUTE", 493, 514), ("CR_S", 514, 532), ("CR_P", 532, 542),
    ("CR_PM", 542, 579), ("CR_PS", 579, 598), ("CR_RTE", 598, 640),
]
# shared comparison field -> (line, window name). Route Suffix reads/boxes the
# LOCATION field (the suffix rides the route token there). Note Xing Line Lgth:
# the TSN print carries X_CROSS_OVERRIDE on LINE 1 where TSMIS prints it on the
# record's second line — each side boxes its own layout.
TSN_CELL = {
    "PR": (1, "PP"), "Route Suffix": (1, "LOC"),
    "District": (1, "LOC"), "County": (1, "LOC"),
    "Date of Record": (1, "DATE_REC"),
    "HG": (1, "HG"), "City Code": (1, "CITY"), "R/U": (1, "RU"),
    "INT Type Eff-Date": (1, "EFF_DATE_INT"), "INT Type": (1, "TY_INT"),
    "Control Type Eff-Date": (1, "EFF_DATE_CT"), "Control Type": (1, "TY_CT"),
    "Lighting Eff-Date": (1, "EFF_DATE_LT"), "Lighting": (1, "LT_TY"),
    "ML Eff-Date": (1, "EFF_DATE_ML"), "ML Mastarm": (1, "MAIN_SM"),
    "ML Left Chan": (1, "MAIN_LC"), "ML Right Chan": (1, "MAIN_RC"),
    "ML Traffic Flow": (1, "MAIN_TF"), "ML Num Lanes": (1, "MAIN_NL"),
    "Xing Line Lgth": (1, "X_CROSS_OVERRIDE"),
    "Description": (2, "DESCRIPTION"), "Main Line Length": (2, "MAIN_OVERRIDE"),
    "CS Eff-Date": (2, "CROSS_BEGIN_DATE"), "CS Mastarm": (2, "CS_SM"),
    "CS Left Chan": (2, "CS_LC"), "CS Right Chan": (2, "CS_RC"),
    "CS Traffic Flow": (2, "CS_TF"), "CS Num Lanes": (2, "CS_NL"),
    "Int St Eff-Date": (2, "EFF_DATE"),
    "Intrte Route": (2, "CR_ROUTE"), "Intrte PM Prefix": (2, "CR_P"),
    "Intrte Postmile": (2, "CR_PM"), "Intrte PM Suffix": (2, "CR_PS"),
}
_PM_LINE_RE = re.compile(r"^\d{3}\.\d{3}$")
_LOC_RE = re.compile(r"^(\d{2})\s+([A-Z]{1,4}\.?)\s+(\d{3}[A-Z]?)$")
# A TASAS date may carry a glued signature flag ('Y98-08-28' / '*98-08-28');
# the flag isn't part of the stored value.
_FLAGGED_DATE_RE = re.compile(r"^[*Y](?=\d{2}-\d{2}-\d{2}$)")

_WORD_GAP = 1.6
# One statewide print is ~1,100 pages of word extraction — index each file ONCE
# and serve every district's locate from the cache (keyed on size+mtime; a few
# entries so per-district drops work too).
_INDEX_CACHE = {}
_INDEX_CACHE_MAX = 4


def _words_of(chars):
    ws, cur, last = [], None, None
    for ch in sorted(chars, key=lambda c: c["x0"]):
        if last is None or ch["x0"] - last > _WORD_GAP:
            if cur:
                ws.append(cur)
            cur = {"t": ch["text"], "x0": ch["x0"], "x1": ch["x1"]}
        else:
            cur["t"] += ch["text"]
            cur["x1"] = ch["x1"]
        last = ch["x1"]
    if cur:
        ws.append(cur)
    return ws


def _assign_win(words, windows):
    """{window name: (joined text, [word boxes])} by MAX x-overlap per word."""
    hit = {name: [] for name, _lo, _hi in windows}
    for w in words:
        best, best_ov = None, 0.0
        for name, lo, hi in windows:
            ov = min(w["x1"], hi) - max(w["x0"], lo)
            if ov > best_ov:
                best, best_ov = name, ov
        if best is not None:
            hit[best].append(w)
    return {name: (" ".join(w["t"] for w in ws), ws) for name, ws in hit.items()}


def _line_rec(page_no, chars):
    words = _words_of(chars)
    return {"page": page_no, "words": words,
            "top": min(c["top"] for c in chars),
            "bottom": max(c["bottom"] for c in chars)}


def _print_index(path, events=None):
    """Parse one TSN Intersection Detail print into {(county, route, pm):
    [record]} + the set of districts it covers. Cached on the print's CONTENT
    identity — the statewide file is indexed once, then every district's locate is a lookup."""
    path = Path(path)
    # CMP-AUD-080: keyed on the print's CONTENT identity, not (size, mtime). A
    # same-size, timestamp-restored replacement used to return the previous
    # parse without reopening the file.
    sig = artifact_store.content_digest(path)
    ent = _INDEX_CACHE.get(str(path))
    if ent and ent["sig"] == sig:
        return ent
    records = defaultdict(list)
    districts = set()
    pending = None
    with pdfplumber.open(path) as pdf:
        n_pages = len(pdf.pages)
        for pi, page in enumerate(pdf.pages, 1):
            if events is not None and pi % 200 == 0:
                events.on_log(f"    …TSN print {path.name}: page {pi}/{n_pages}")
            page_chars = [c for c in page.chars if c["text"].strip()
                          and c.get("fontname", "").endswith("Courier")]
            # a statewide print is ~1,100 pages in ONE handle — without this,
            # pdfplumber's per-page caches accumulate to gigabytes by the end.
            page.flush_cache()
            for _top, chars in cluster_by_top(page_chars, 3):
                rec = _line_rec(pi, chars)
                a1 = _assign_win(rec["words"], _L1_WIN)
                m_loc = _LOC_RE.match(a1["LOC"][0])
                if _PM_LINE_RE.match(a1["PM"][0]) and m_loc:
                    pending = (rec, a1, m_loc)
                    continue
                if pending is None:
                    continue
                l1, a1p, m_loc = pending
                pending = None
                a2 = _assign_win(rec["words"], _L2_WIN)
                dist, cnty = m_loc.group(1), m_loc.group(2).rstrip(".")
                base, _sfx = idt._split_route(a1p["LOC"][0])
                districts.add(dist)
                key = (cnty, base, idt._norm_pm(a1p["PM"][0]))
                records[key].append({"l1": l1, "a1": a1p, "l2": rec, "a2": a2,
                                     "dist": dist})
    ent = {"sig": sig, "districts": districts, "records": dict(records)}
    while len(_INDEX_CACHE) >= _INDEX_CACHE_MAX:
        _INDEX_CACHE.pop(next(iter(_INDEX_CACHE)))
    _INDEX_CACHE[str(path)] = ent
    return ent


def district_index(pdf_dir, events=None):
    """{district('01'..'12'): path} by INDEXING each PDF and reading which
    districts its own records cover (filenames are the user's business; the
    single statewide print maps every district to itself). The heavy word
    extraction happens here, once per file — locate_tsn is lookups after."""
    index = {}
    for p in sorted(Path(pdf_dir).glob("*.pdf")):
        try:
            ent = _print_index(p, events)
        except Exception as e:                            # unreadable/odd PDF
            log.warning("evidence: %s unreadable: %s: %s",
                        p.name, type(e).__name__, e)
            if events:
                events.on_log(f"    note: {p.name} unreadable, skipped")
            continue
        if not ent["districts"]:
            log.info("evidence: %s has no Intersection Detail records; skipped",
                     p.name)
            continue
        for dist in ent["districts"]:
            index.setdefault(dist, p)
    return index


def locate_tsn(pdf_path, needed_routes, needed_keys):
    """{(county, route, pm): [record]} served from the file's cached index."""
    del needed_routes                       # the key filter is already exact
    ent = _print_index(pdf_path)
    return {k: ent["records"][k] for k in needed_keys if k in ent["records"]}


def _tsn_raw(rec, field):
    """The raw print token for `field` (signature flags stripped from dates)."""
    if field == "Route Suffix":
        return idt._split_route(rec["a1"]["LOC"][0])[1]
    if field in ("District", "County"):
        district, county = idt._dist_cnty(rec["a1"]["LOC"][0])
        return district if field == "District" else county
    line, name = TSN_CELL[field]
    text = (rec["a1"] if line == 1 else rec["a2"])[name][0]
    return _FLAGGED_DATE_RE.sub("", text)


def tsn_value(rec, field):
    """The compared value this print record carries for `field`."""
    if field in ("Route Suffix", "District", "County"):
        return _tsn_raw(rec, field)
    return project(field, _tsn_raw(rec, field))


def tsn_box(rec, field):
    """(page_no, cell_box, record_yspan, record_xspan) for `field`'s cell. A
    blank cell boxes its fixed template window — on a line-printer template the
    window IS the cell."""
    line, name = TSN_CELL[field]
    lrec = rec["l1"] if line == 1 else rec["l2"]
    assigned = (rec["a1"] if line == 1 else rec["a2"])[name][1]
    if assigned:
        x0, x1 = min(w["x0"] for w in assigned), max(w["x1"] for w in assigned)
    else:
        windows = _L1_WIN if line == 1 else _L2_WIN
        lo, hi = next((lo, hi) for n, lo, hi in windows if n == name)
        x0, x1 = lo + 1, hi - 3
    words = rec["l1"]["words"] + rec["l2"]["words"]
    xspan = (min(w["x0"] for w in words) - 4, max(w["x1"] for w in words) + 4)
    yspan = ((rec["l1"]["top"], rec["l2"]["bottom"])
             if rec["l1"]["page"] == rec["l2"]["page"]
             else (lrec["top"], lrec["bottom"]))
    return (lrec["page"], (x0 - 2, lrec["top"] - 2, x1 + 2, lrec["bottom"] + 2),
            yspan, xspan)
