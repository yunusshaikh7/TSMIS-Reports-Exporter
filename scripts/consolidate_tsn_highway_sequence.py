"""Convert the TSN district Highway Sequence (HSL) PDFs into one normalized
Excel the Highway Sequence comparison reads.

The TSN "Highway Locations" report (OTM22025) is a fixed-layout PDF, one per
district (D01..D12), listing every route's postmile sequence: a per-page header
band, a centered "DIST NN RTE NNN DIR X-X" group header, then one data line per
location:

    CO.  CITY  POSTMILE  G/RF  DISTANCE-TO-NXT-POINT  DESCRIPTION
    MEN        000.000    UH   000.056               BEGIN OF COUNTY
    ORA  DAPT  R000.129   DH   000.102               JCT 5 CAMINO L RMBLS UC

Reconciled by hand against the per-route TSMIS "Highway Locations" XLSX
(docs/tsn-parsers.md):
  * CALIFORNIA postmiles are COUNTY-RELATIVE — the same route restarts at
    000.000 in each county it crosses — so the comparison keys on
    route + county + postmile, not route + postmile.
  * The POSTMILE token carries a glued realignment prefix ("R000.129") and/or
    an equate suffix ("050.025E"); TSMIS stores those in separate (unnamed)
    prefix/suffix columns. We keep the glued TSN form as the canonical postmile
    and the TSMIS loader re-glues prefix+PM+suffix to match.
  * The 2-letter "G/RF" flag is HG (1st char) + FT (2nd char) — TSMIS splits
    them into its HG and FT columns.
  * TSN lists every segment break (incl. unnamed ones) and prints equate points
    as a "Rxxx EQUATES TO" annotation; TSMIS omits most unnamed breaks and
    stores the equate as an "END R REALIGNMENT" row. Both differences surface
    honestly as one-sided rows in the comparison (as they already do for the
    Highway Log — "mostly TSN segment splits and TSMIS realignment markers").

Normalization v4 is SOURCE-EXACT (CMP-AUD-155/156/158/159): the printed
DISTANCE TO NXT POINT pointer markers ("*P*" / "-------->") are conserved
verbatim and any other non-numeric token there refuses; the 46 statewide
annotations printed before their route's first county-bearing row are conserved
with their BLANK county (never dropped or backfilled — the cover itself warns
equate ownership may be wrong); wrapped descriptions join on a single space (no
invented punctuation); and every document's identity claims (cover NOTE policy,
report id/dates/times exactly-once, per-route printed directions) are captured,
cross-member-checked, and persisted via producer_extra → the library sidecar.
The workbook carries a "TSN Normalization" marker sheet the comparison loader
gates on (the rows sheet kept its shape, so the marker is the version signal).

Unlike word-fusion-prone Highway Log, HSL columns are widely spaced, so
word-level extraction (with the 2-char flag split) is safe; only the postmile
token must be classed by its x-window. Columns are calibrated to the OTM22025
layout and verified stable across all 12 districts.

Console-free like the other consolidators: progress via events.on_log, overwrite
through the confirm_overwrite callback, cancel honored between pages, a
ConsolidateResult returned. The console UX lives in cli.run_consolidate_cli.
"""
import hashlib
import io
import logging
import re
from pathlib import Path

logging.getLogger("pdfminer").setLevel(logging.ERROR)

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from compare_core import is_formula_injection
from events import ConsolidateResult, Events
from pdf_table_lib import cluster_by_top, norm_route
import outcome
import artifact_store
import tsn_district_contract as tdc
from paths import OUTPUT_ROOT, tsn_library_raw_dir

# Standalone default locations — see the matching note in
# consolidate_tsn_highway_log.py. Reads the district PDFs from the canonical TSN
# library raw folder (v0.30.0 retired the separate input/ folder); the
# GUI/matrices build through the same tsn_library/ tree. This OUT_PATH holds
# Caltrans-internal TSN data and is git-ignored (output/* + the explicit
# output/tsn_* rule in .gitignore) — never add an "!output/tsn_*" allowlist entry.
INPUT_DIR = tsn_library_raw_dir("highway_sequence")
OUT_PATH = OUTPUT_ROOT / "tsn_highway_sequence_consolidated.xlsx"

REPORT_NAME = "TSN Highway Sequence"
INPUT_GLOB = "*.pdf"
INPUT_FMT = "PDF"
INPUT_NOTE = "Drop the TSN district Highway Sequence (HSL) PDFs into the TSN library folder first."

# The normalized workbook the comparison reads: a leading Route column + the
# shared comparison fields, in this exact order (compare_highway_sequence_tsn
# reads it positionally). Postmile is the glued canonical form (prefix+PM+suffix).
NORMALIZED_SHEET = "Highway Locations (TSN)"
NORMALIZED_HEADER = ["Route", "County", "PM", "City", "HG", "FT",
                     "Distance To Next Point", "Description"]

# v4 (CMP-AUD-155/156/158/159): the row universe and values are source-exact —
# printed pointer tokens kept, pre-county equates kept (blank County), wrapped
# descriptions joined without invented punctuation — and the print's identity/
# direction/policy claims ride the sidecar. The rows sheet is unchanged in
# SHAPE, so the loader detects an old workbook via this marker sheet instead
# (report_catalog's TsnEntry version must match; check_compare_highway_sequence_tsn
# asserts the pair).
NORMALIZATION_VERSION = 4
MARKER_SHEET = "TSN Normalization"


# =============================================================================
# PDF layout — calibrated to OTM22025 "Highway Locations" (verified D01..D12)
# =============================================================================

Y_TOLERANCE = 3      # words within this y-distance form one logical line

# A word belongs to the column whose x-window contains its left edge (x0). The
# columns are widely separated (no fusion risk), so x0 classification is stable.
# The 2-char G/RF flag is one fused token ("UH") split into HG + FT below.
W_COUNTY = (0, 44)
W_CITY = (44, 98)
W_PM = (98, 168)
W_FLAG = (168, 205)
W_DIST = (205, 270)
W_DESC = (270, 700)

# A postmile, optionally with a glued realignment prefix ("R012.887") and/or an
# equate suffix ("050.025E"), exactly as printed in the POSTMILE column.
LOCATION_RE = re.compile(r"^[A-Z]?\d{3}\.\d{3}[A-Z]?$")
# A county code in the CO. column ("MEN", "LA.", "SB.").
COUNTY_RE = re.compile(r"^[A-Z]{2,4}\.?$")
# A distance-to-next value ("000.056"); TSN prints 3-int-digit zero-padded.
DIST_RE = re.compile(r"^\d{1,3}\.\d{3}$")
# The print's two non-numeric DISTANCE TO NXT POINT claims (CMP-AUD-156): a
# point/pointer marker occupying the distance data position. Censused across
# all 12 districts / 1,540 pages: exactly these two forms (283 "*P*" +
# 282 "-------->" = 565). Conserved verbatim; any OTHER token in the distance
# window refuses loudly (layout drift needs a re-census, not a silent blank).
POINTER_TOKENS = ("*P*", "-------->")
# Centered group header: "DIST 01 RTE 001 DIR S-N" (route may be suffixed:
# 005S). The direction is a per-route source claim (CMP-AUD-155) — captured,
# consistency-checked, and persisted with the sidecar claims.
GROUP_RE = re.compile(r"\bDIST\s+(\d{1,2})\s+RTE\s+([0-9A-Z]+)\s+DIR\s+([NSEW]-[NSEW])\b")
GROUP_LIKE_RE = re.compile(r"\bDIST\b.*\bRTE\b.*\bDIR\b", re.IGNORECASE)

# CMP-AUD-155 — the print's identity claims. Every data page carries a header
# band (the lines above and including the DIST/RTE/DIR line): the report id,
# report date, title, "Ref Dt", and the generation time; the cover (page 1)
# carries the report id, "Reference Date:", "District:", and the TSN
# reliability NOTE. Each field must resolve to exactly ONE distinct value per
# document; report identity must also agree across the 12 districts.
REPORT_ID_RE = re.compile(r"\bOTM\d+\b")
REPORT_TITLE_RE = re.compile(r"\bHighway\s+Locations\b")
REPORT_DATE_RE = re.compile(r"\b\d{2}-[A-Z]{3}-\d{2}\b")
REF_DT_RE = re.compile(r"\bRef\s+Dt\s+(\d{2}\s+[A-Z]{3}\s+\d{4})\b")
GEN_TIME_RE = re.compile(r"\b\d{2}:\d{2}\s+[AP]M\b")
COVER_REF_DATE_RE = re.compile(r"Reference\s+Date:\s*([0-9A-Z-]+)")
COVER_DISTRICT_RE = re.compile(r"District:\s*(\d{1,2})\b")
POLICY_MARKER = "* * * N O T E * * *"


def _bucket(x0):
    for name, (lo, hi) in (("county", W_COUNTY), ("city", W_CITY), ("pm", W_PM),
                           ("flag", W_FLAG), ("dist", W_DIST), ("desc", W_DESC)):
        if lo <= x0 < hi:
            return name
    return None


def _cluster_lines(words):
    """Group words into logical lines, tolerating the ~1pt baseline jitter of the
    proportional font (a plain round(top) splits a single row across two buckets
    when its words straddle x.5, dropping rows that then lack a county or PM)."""
    return [ws for _top, ws in cluster_by_top(words, Y_TOLERANCE)]


# The canonical route-token normalizer (pdf_table_lib; behavior unchanged —
# this module's copy was already the reconciled regex form).
_norm_route = norm_route


def _parse_line(ws):
    """Classify one clustered line's words into column buckets -> {bucket: [text]}."""
    cols = {}
    for w in ws:
        b = _bucket(w["x0"])
        if b is not None:
            cols.setdefault(b, []).append(w["text"])
    return cols


def _single_claim(pdf_name, name, values):
    """CMP-AUD-155 multiplicity rule: a printed identity field may appear on
    many pages but must resolve to exactly ONE distinct value per document;
    conflicting or missing values refuse (re-census the print before trusting)."""
    distinct = sorted(set(values))
    if len(distinct) != 1:
        seen = distinct if distinct else "no value at all"
        raise ValueError(
            f"{pdf_name}: the print's {name} must appear with exactly one "
            f"distinct value across the document; saw {seen}")
    return distinct[0]


def _parse_cover(page, pdf_name):
    """Page 1 of every district print is the cover: the report id, the
    'Reference Date:' line, the 'District:' line, and the TSN reliability NOTE
    warning that equate/boundary descriptions may be wrong (CMP-AUD-155). Every
    role is required — a cover missing them is not a TSN Highway Sequence print."""
    text = page.extract_text() or ""
    flat = " ".join(text.split())
    start = flat.find(POLICY_MARKER)
    if start < 0:
        raise ValueError(
            f"{pdf_name}: the cover's TSN reliability NOTE ({POLICY_MARKER!r}) "
            "is missing — not a TSN Highway Sequence cover page")
    policy = flat[start:]
    return {
        "report_id": _single_claim(
            pdf_name, "cover report id", REPORT_ID_RE.findall(flat)),
        "cover_reference_date": _single_claim(
            pdf_name, "cover Reference Date", COVER_REF_DATE_RE.findall(flat)),
        "cover_district": _single_claim(
            pdf_name, "cover District", COVER_DISTRICT_RE.findall(flat)),
        "policy_text": policy,
        "policy_sha256": hashlib.sha256(policy.encode("utf-8")).hexdigest(),
    }


def parse_pdf(path, events, pdf_name=""):
    """Parse one HSL PDF -> ``(internal district, routes, claims)`` in document
    order. `claims` is the document's CMP-AUD-155 source-claims record: report
    identity/timing (exactly one distinct value each), the cover's reliability
    NOTE, and the per-route printed direction."""
    routes = {}
    district_claims = []
    directions = {}            # route -> printed DIR ("S-N"…); conflict refuses
    band = {"report id": [], "report title": [], "report date": [],
            "reference date (Ref Dt)": [], "generation time": []}
    route = None
    county = None              # carried onto equate-annotation rows (no county token)
    last_row = None            # a wrapped description line attaches here

    with pdfplumber.open(path) as pdf:
        n_pages = len(pdf.pages)
        if not n_pages:
            raise ValueError(f"{pdf_name}: the PDF has no pages")
        cover = _parse_cover(pdf.pages[0], pdf_name)
        district_claims.append(cover["cover_district"])
        for page_no, page in enumerate(pdf.pages[1:], 2):
            if events.is_cancelled():
                return None, None, None
            if page_no % 25 == 0:
                events.on_log(f"    …page {page_no}/{n_pages}")
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            lines = _cluster_lines(words)
            texts = [" ".join(w["text"] for w in ws) for ws in lines]

            # CMP-AUD-155: the page's identity band = every line above and
            # including the DIST/RTE/DIR group header. Collect the printed
            # identity fields; each must resolve to ONE distinct value.
            group_idx = next(
                (i for i, t in enumerate(texts) if GROUP_RE.search(t)), None)
            if group_idx is not None:
                band_text = " ".join(texts[:group_idx + 1])
                band["report id"] += REPORT_ID_RE.findall(band_text)
                band["report title"] += REPORT_TITLE_RE.findall(band_text)
                band["report date"] += REPORT_DATE_RE.findall(band_text)
                band["reference date (Ref Dt)"] += REF_DT_RE.findall(band_text)
                band["generation time"] += GEN_TIME_RE.findall(band_text)

            for line, text in zip(lines, texts):
                gm = GROUP_RE.search(text)
                if gm:
                    district_claims.append(gm.group(1))
                    route = _norm_route(gm.group(2))
                    routes.setdefault(route, [])
                    # The printed direction is a per-route source claim
                    # (CMP-AUD-155); a route printing two directions in one
                    # document is a source inconsistency, not a choice.
                    direction = gm.group(3)
                    prev = directions.get(route)
                    if prev is not None and prev != direction:
                        raise ValueError(
                            f"{pdf_name} p{page_no}: route {route} printed "
                            f"conflicting directions {prev!r} and {direction!r}")
                    directions[route] = direction
                    county = None
                    last_row = None
                    continue
                if GROUP_LIKE_RE.search(text):
                    raise ValueError(
                        f"{pdf_name} p{page_no}: malformed DIST/RTE/DIR group header "
                        f"cannot safely own following rows: {text!r}")

                cols = _parse_line(line)
                co = (cols.get("county") or [""])[0]
                pm = next((t for t in (cols.get("pm") or []) if LOCATION_RE.match(t)), None)

                # Equate annotation: "Rxxx.xxx EQUATES TO" — a postmile at the PM
                # window, NO county token. TSMIS records the same equate as an
                # "END R REALIGNMENT" row at that postmile, so emit it (county
                # carried from context) to pair the two rather than orphan it.
                # CMP-AUD-158: an annotation printed BEFORE the route's first
                # county-bearing row is conserved WITH ITS BLANK COUNTY — the
                # source cover itself warns equate descriptions may be wrong, so
                # its unknown ownership is disclosed, never backfilled or dropped.
                if "EQUATES" in text and pm and not COUNTY_RE.match(co):
                    if route is None:
                        raise ValueError(
                            f"{pdf_name} p{page_no}: an EQUATES annotation appeared "
                            "before any owning route header")
                    row = dict(county=county, pm=pm, city=None, hg=None,
                               ft=None, dist=None, description="EQUATES TO")
                    routes[route].append(row)
                    last_row = row
                    continue

                # Data line: a county code AND a postmile in their windows.
                if COUNTY_RE.match(co) and pm:
                    if route is None:
                        raise ValueError(
                            f"{pdf_name} p{page_no}: recognizable highway-sequence "
                            "data appeared before any owning route header")
                    county = co.rstrip(".")
                    flag = "".join(cols.get("flag") or [])
                    desc = " ".join(cols.get("desc") or []).strip() or None
                    # CMP-AUD-156: the DISTANCE TO NXT POINT position carries
                    # either a numeric distance or a printed pointer marker
                    # ("*P*" / "-------->"); both are source claims, conserved
                    # verbatim. Any other token there is layout drift — refuse.
                    dists = cols.get("dist") or []
                    dist = dists[0] if dists else None
                    if dist is not None and not (
                            DIST_RE.match(dist) or dist in POINTER_TOKENS):
                        raise ValueError(
                            f"{pdf_name} p{page_no}: unrecognized DISTANCE TO NXT "
                            f"POINT token {dist!r} — the print layout drifted; "
                            "re-census before trusting this conversion")
                    row = dict(
                        county=county, pm=pm,
                        city=(cols.get("city") or [None])[0],
                        hg=(flag[0] if len(flag) >= 1 else None),
                        ft=(flag[1] if len(flag) >= 2 else None),
                        dist=dist, description=desc)
                    routes[route].append(row)
                    last_row = row
                    continue

                # A wrapped description continuation: text only in the description
                # window, no county/postmile -> append to the open row.
                # CMP-AUD-159: fragments join on a single space — the source
                # prints no punctuation between wrapped lines, so none may be
                # invented.
                if (last_row is not None and cols.get("desc")
                        and not cols.get("county") and not cols.get("pm")):
                    extra = " ".join(cols["desc"]).strip()
                    if extra:
                        last_row["description"] = (
                            extra if not last_row["description"]
                            else last_row["description"] + " " + extra)

    district = tdc.document_district(district_claims, pdf_name or Path(path).name)
    claims = {
        "member": pdf_name or Path(path).name,
        "district": district,
        "report_id": _single_claim(
            pdf_name, "report id", [cover["report_id"], *band["report id"]]),
        "report_title": _single_claim(pdf_name, "report title", band["report title"]),
        "report_date": _single_claim(pdf_name, "report date", band["report date"]),
        "reference_date": _single_claim(
            pdf_name, "reference date (Ref Dt)", band["reference date (Ref Dt)"]),
        "cover_reference_date": cover["cover_reference_date"],
        "generation_time": _single_claim(
            pdf_name, "generation time", band["generation time"]),
        "pages": n_pages,
        "policy_sha256": cover["policy_sha256"],
        "policy_text": cover["policy_text"],
        "directions": directions,
    }
    return district, routes, claims


def _cross_member_claims(document_claims):
    """CMP-AUD-155: the 12 district prints must agree on the report identity —
    a member from a DIFFERENT TSN pull (another report/reference date) silently
    poisoning the library is exactly the cross-member inconsistency this
    refuses. Generation times legitimately differ per district and stay
    per-document; the reliability policy is conserved per document (distinct
    texts recorded, never required identical)."""
    record = {"schema_version": 1}
    for field in ("report_id", "report_title", "report_date",
                  "reference_date", "cover_reference_date"):
        values = sorted({str(d[field]) for d in document_claims})
        if len(values) != 1:
            members = {d["member"]: str(d[field]) for d in document_claims}
            raise ValueError(
                f"the district prints disagree on the {field.replace('_', ' ')} "
                f"({members}) — all 12 must come from the same TSN pull")
        record[field] = values[0]
    record["documents"] = [
        {"member": d["member"], "district": d["district"],
         "generation_time": d["generation_time"], "pages": d["pages"],
         "policy_sha256": d["policy_sha256"],
         "directions": dict(sorted(d["directions"].items()))}
        for d in sorted(document_claims, key=lambda d: d["district"])
    ]
    record["policy_texts"] = sorted({d["policy_text"] for d in document_claims})
    return record


# =============================================================================
# Normalized workbook
# =============================================================================

def _row_values(route, d):
    """One normalized row in NORMALIZED_HEADER order."""
    return [route, d["county"], d["pm"], d["city"], d["hg"], d["ft"],
            d["dist"], d["description"]]


def _write_workbook(all_rows, out_path, proceed=None):
    """Write every (route, row_dict) into one normalized workbook (write_only so
    the statewide ~46k rows never exhaust memory)."""
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(NORMALIZED_SHEET)
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 9
    for i, name in enumerate(NORMALIZED_HEADER[1:], start=2):
        ws.column_dimensions[get_column_letter(i)].width = \
            40 if name == "Description" else 14

    head = []
    for name in NORMALIZED_HEADER:
        c = WriteOnlyCell(ws, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        head.append(c)
    ws.append(head)

    for route, d in all_rows:
        cells = []
        for v in _row_values(route, d):
            if is_formula_injection(v):       # neutralize a "="-leading description
                c = WriteOnlyCell(ws, value=v)
                c.data_type = "s"
                cells.append(c)
            else:
                cells.append(v)
        ws.append(cells)

    # The v4 shape marker (CMP-AUD-156/158/159): the rows sheet is unchanged in
    # WIDTH, so a bare pre-v4 workbook is indistinguishable by shape — this
    # sheet is what the comparison loader gates on (the library path already
    # auto-rebuilds via report_catalog's normalization_version, D2).
    marker = wb.create_sheet(MARKER_SHEET)
    marker.append(["Report", REPORT_NAME])
    marker.append(["Normalization version", NORMALIZATION_VERSION])
    # F9 temp + os.replace + the P12 TOCTOU gate at the replace.
    return artifact_store.atomic_save_if(wb, out_path, proceed or (lambda: True))


# =============================================================================
# Entry points
# =============================================================================

def build_into(raw_dir, out_path, events=None, confirm_overwrite=None):
    """Canonical-TSN-library entry point: parse the district PDFs in `raw_dir`,
    write the one normalized workbook to `out_path`. Thin wrapper over
    consolidate() so tsn_library drives every report through one signature."""
    return consolidate(events=events, confirm_overwrite=confirm_overwrite,
                       input_dir=raw_dir, out_path=out_path)


def consolidate(events=None, confirm_overwrite=None, day=None,
                input_dir=None, out_path=None):
    """Parse every TSN district HSL PDF in `input_dir` into one normalized
    workbook at `out_path`. `day` is accepted for interface symmetry and ignored
    (TSN PDFs are vendor snapshots, not dated exports). Console-free."""
    in_dir = Path(input_dir) if input_dir else INPUT_DIR
    out = Path(out_path) if out_path else OUT_PATH
    events = events or Events()
    if not _DEPS_OK:
        return ConsolidateResult(
            status="error",
            message="Required components are missing (pdfplumber, openpyxl).")
    confirm = confirm_overwrite or (lambda _p: True)

    try:
        in_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass

    pdfs = sorted(p for p in in_dir.glob("*.pdf") if p.is_file())
    if not pdfs:
        return ConsolidateResult(
            status="error",
            message=(f"No {REPORT_NAME} files were found in:\n{in_dir}\n\n"
                     f"Put the district Highway Sequence PDFs (e.g. "
                     f"D01 HSL TSN.pdf) there, then run again."))

    existed_at_confirm = out.exists()
    if existed_at_confirm and not confirm(out):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"TSN Highway Sequence Conversion - {len(pdfs)} district PDF(s)")
    events.on_log("=" * 60)
    events.on_log("")

    try:
        source_manifest, source_bytes = tdc.capture_raw_manifest(pdfs, in_dir)
    except ValueError as e:
        return ConsolidateResult(
            status="error",
            message=f"The TSN Highway Sequence raw source could not be bound: {e}",
        )

    source_problem = [None]

    def source_current():
        try:
            current_pdfs = sorted(p for p in in_dir.glob("*.pdf") if p.is_file())
            current = tdc.canonical_raw_manifest(current_pdfs, in_dir)
        except (OSError, ValueError) as e:
            source_problem[0] = f"{type(e).__name__}: {e}"
            return False
        if current != source_manifest:
            source_problem[0] = "the raw member names or bytes changed during conversion"
            return False
        return True

    def source_changed():
        detail = source_problem[0] or "the raw source changed during conversion"
        return ConsolidateResult(
            status="error",
            message=(f"The TSN Highway Sequence raw source changed while it was being "
                     f"normalized ({detail}); last-good output was preserved."),
        )

    # route -> [row_dict]; same route may appear across districts (different
    # counties) — accumulate so all of a route's counties land in one workbook.
    by_route = {}
    failed = []
    claimed_districts = {}
    document_claims = []
    for i, p in enumerate(pdfs, 1):
        if events.is_cancelled():
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i}/{len(pdfs)}] {p.name}"
        events.on_log(f"{prefix} parsing…")
        try:
            district, route_rows, doc_claims = parse_pdf(
                io.BytesIO(source_bytes[p.name]), events, pdf_name=p.name)
        except Exception as e:
            events.on_log(f"{prefix} FAILED ({type(e).__name__}): {e}")
            failed.append(p.name)
            continue
        if route_rows is None:
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        if not route_rows:
            events.on_log(f"{prefix} no highway-sequence data found; skipping")
            failed.append(p.name)
            continue
        if district in claimed_districts:
            return ConsolidateResult(
                status="error", failed_inputs=1,
                message=(f"Duplicate internal district claim D{district}: "
                         f"{claimed_districts[district]} and {p.name}. "
                         "Nothing was published."),
            )
        claimed_districts[district] = p.name
        document_claims.append(doc_claims)
        n = 0
        for route, rows in route_rows.items():
            by_route.setdefault(route, []).extend(rows)
            n += len(rows)
        events.on_log(f"{prefix} +{n} rows across {len(route_rows)} route(s)")

    if not by_route:
        return ConsolidateResult(
            status="error",
            message=(f"None of the PDFs in:\n{in_dir}\n\ncontained readable "
                         f"{REPORT_NAME} data. Are they the TSN Highway Locations PDFs?"))

    if failed:
        return ConsolidateResult(
            status="error", failed_inputs=len(failed),
            message=(f"The TSN Highway Sequence source is incomplete or unreadable: "
                     f"{len(failed)} document(s) failed {failed}. Exactly one "
                     "internally claimed document for every D01-D12 is required; "
                     "nothing was published."),
        )
    try:
        tdc.require_exact_universe(
            (district, claimed_districts[district])
            for district in claimed_districts)
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))
    try:
        source_claims = _cross_member_claims(document_claims)
    except ValueError as e:
        return ConsolidateResult(
            status="error",
            message=f"The TSN Highway Sequence source claims are inconsistent: {e}")
    if events.is_cancelled():
        return ConsolidateResult(status="cancelled", message="Cancelled by user.")
    if not source_current():
        return source_changed()

    all_rows = [(route, d) for route in sorted(by_route)
                for d in by_route[route]]

    out.parent.mkdir(parents=True, exist_ok=True)
    events.on_log("")
    events.on_log("Writing normalized workbook...")
    final_gate = [None]

    def may_publish():
        if events.is_cancelled():
            final_gate[0] = "cancelled"
            return False
        if not source_current():
            final_gate[0] = "source"
            return False
        if not artifact_store.confirm_late_overwrite(out, existed_at_confirm, confirm):
            final_gate[0] = "overwrite"
            return False
        return True

    try:
        # P12 TOCTOU: the overwrite gate is INSIDE _write_workbook, at the os.replace
        # (atomic_save_if) — a destination that appears during the BUILD is caught.
        committed = _write_workbook(all_rows, out, proceed=may_publish)
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not save {out.name}.\n\n"
                     "The file is probably open in Excel. Close it and try again."))
    if not committed:
        if final_gate[0] == "source":
            return source_changed()
        if final_gate[0] == "cancelled":
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.")

    # CMP-AUD-035: re-verify the raw source AFTER the os.replace. may_publish()
    # checks source_current() just BEFORE the replace, but a change in that window
    # would still publish bytes built from the now-stale source and return success.
    # The canonical build_consolidated wrapper rehashes post-builder; this
    # DIRECT/CLI path needs its own post-replace recheck so it never returns a
    # success-shaped result for a source that changed during the final commit.
    if not source_current():
        return ConsolidateResult(
            status="error",
            message=("The TSN Highway Sequence raw source changed during the final "
                     f"commit ({source_problem[0] or 'raw member names or bytes changed'})"
                     "; re-run to normalize the current source."))

    summary_lines = []
    summary_lines += [
        f"District PDFs:  {len(pdfs)} parsed; exact D01-D12",
        f"Routes:         {len(by_route)}",
        f"Rows:           {len(all_rows)}",
        f"Print identity: {source_claims['report_id']} "
        f"{source_claims['report_title']} · report {source_claims['report_date']} "
        f"· reference {source_claims['reference_date']}",
        f"Output file:    {out}",
    ]
    result = ConsolidateResult(status="ok", output_path=str(out),
                               summary_lines=summary_lines,
                               completion=outcome.COMPLETE, failed_inputs=0)
    result.tsn_raw_manifest = source_manifest
    # CMP-AUD-155: the print's identity/direction/policy claims ride the
    # library sidecar (tsn_library.build_normalized merges producer_extra).
    result.producer_extra = {"tsn_source_claims": source_claims}
    return result


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
