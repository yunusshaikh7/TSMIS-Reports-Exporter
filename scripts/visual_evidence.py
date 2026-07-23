"""Visual evidence for vs-TSN comparisons: highlighted PDF snippets per diff.

For each shared column the comparison flags somewhere, sample random example
rows (random routes — not just the first), find that exact cell in BOTH source
PDFs (the TSMIS per-route "(PDF)" export and the TSN district print), render
the region and box the cell on each side, and compose captioned evidence
images — the manual screenshot-and-circle workflow, automated.

Trust contract: an example is only used when the value parsed back OUT of each
PDF, normalized with the comparator's own projections, equals the value the
comparison actually compared — so an image can never illustrate something other
than what was diffed, and every rendered pair doubles as an end-to-end check of
the comparison at that cell. Candidates that fail (the TSMIS PDF/Excel editions
disagreeing at that cell, a TSN reference-date skew, a duplicated key) are
skipped with a recorded reason, never shown.

Outputs, next to the comparison workbook (the "(formulas).xlsx" sibling
convention): `<comparison> (evidence).xlsx` — a Summary sheet + one tab PER
COMPARISON COLUMN for the chosen layout — and `<comparison> (evidence images)/`
holding the same examples as loose files. The user picks the layout in Settings
(side-by-side for pasting into docs, stacked for reading, or both — 'both' gives
each column two tabs and writes both PNGs); only the selected layout(s) are
rendered. Both writes are keep-last-good: a failed/cancelled run leaves the
previous set untouched; files locked open in Excel divert to an unpredictable
".new-<token>" sibling with a note.

Report-agnostic: everything report-specific comes from an adapter module
(evidence_highway_detail — district TSN prints; evidence_intersection_detail —
the statewide TSN print; each covers its report's Excel + PDF rows). Engine is
console-free (Events sink, cancellation honored between steps) and never
affects the comparison result it decorates.
"""
import collections
import hashlib
import logging
import os
import random
import re
import secrets
import shutil
import tempfile
from pathlib import Path

try:
    import pdfplumber
    from PIL import Image, ImageDraw, ImageFont
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import artifact_store
import evidence_ledger
import evidence_manifest
import owned_dir
import paths
from compare_core import set_safe_literal_cell
from pdf_table_lib import RouteIdentityError

log = logging.getLogger("tsmis.evidence")

DEPS_MSG = "Required components are missing (pdfplumber/Pillow/openpyxl)."

# row_key -> adapter module name (lazy import; each report's Excel + PDF rows
# share one adapter — the Excel row's images render from the PDF-edition export,
# the PDF row's from the same files it was compared from).
_ADAPTER_MODULES = {
    "highway_detail": "evidence_highway_detail",
    "highway_detail_pdf": "evidence_highway_detail",
    "intersection_detail": "evidence_intersection_detail",
    "intersection_detail_pdf": "evidence_intersection_detail",
    "highway_log": "evidence_highway_log",
    "highway_log_pdf": "evidence_highway_log",
    "highway_sequence": "evidence_highway_sequence",
    "highway_sequence_pdf": "evidence_highway_sequence",
    "ramp_detail": "evidence_ramp_detail",
    "ramp_detail_pdf": "evidence_ramp_detail",
}
# Where each row's TSMIS-side PDFs live (the per-route export subdir) and which
# TSN library report holds the TSN prints (per-district files for Highway
# Detail/Log; the single statewide TASAS print for Intersection Detail — the
# adapters read district/county per record, so filenames never matter).
TSMIS_PDF_SUBDIR = {"highway_detail": "highway_detail_pdf",
                    "highway_detail_pdf": "highway_detail_pdf",
                    "intersection_detail": "intersection_detail_pdf",
                    "intersection_detail_pdf": "intersection_detail_pdf",
                    "highway_log": "highway_log_pdf",
                    "highway_log_pdf": "highway_log_pdf",
                    "highway_sequence": "highway_sequence_pdf",
                    "highway_sequence_pdf": "highway_sequence_pdf",
                    "ramp_detail": "ramp_detail_pdf",
                    "ramp_detail_pdf": "ramp_detail_pdf"}
TSN_PDF_REPORT = {"highway_detail": "highway_detail",
                  "highway_detail_pdf": "highway_detail",
                  "intersection_detail": "intersection_detail",
                  "intersection_detail_pdf": "intersection_detail",
                  "highway_log": "highway_log",
                  "highway_log_pdf": "highway_log",
                  "highway_sequence": "highway_sequence",
                  "highway_sequence_pdf": "highway_sequence",
                  "ramp_detail": "ramp_detail",
                  "ramp_detail_pdf": "ramp_detail"}
# Report labels for the availability probe (static so the probe never has to
# import an adapter — a state push must stay cheap; check_visual_evidence pins
# these maps against report_catalog so they can't drift).
_TSN_PDF_LABELS = {"highway_detail": "Highway Detail",
                   "intersection_detail": "Intersection Detail",
                   "highway_log": "Highway Log",
                   "highway_sequence": "Highway Sequence",
                   "ramp_detail": "Ramp Detail"}
# Reports whose TSN prints ARE the library's raw inputs (district-PDF-sourced
# TSN libraries — Highway Log and Highway Sequence): evidence reads the SAME
# files from raw/, so a user with a working vs-TSN comparison already has the
# prints in place. The statewide-XLSX-sourced reports keep the separate
# OPTIONAL pdf/ drop folder.
_TSN_PDFS_IN_RAW = frozenset({"highway_log", "highway_sequence"})

MIN_EXAMPLES, MAX_EXAMPLES, DEFAULT_EXAMPLES = 1, 10, 2
_RES = 180                     # render DPI (points * _RES/72 = pixels)
_SC = _RES / 72.0
_STACK_W = 1900                # stacked strip width (px)
_PAIR_SIDE_W = 1300            # side-by-side per-side width (px)
_CTX_PT = 27                   # vertical context around the record (points)
_PAGE_CACHE_MAX = 16
_EMBED_W = 1500                # embedded stacked-image display width in the workbook
_EMBED_W_PAIR = 2000           # side-by-side embeds are twice as wide — give them more
_PX_PER_ROW = 20               # Excel's default row height in pixels

# The workbook image layout the user picks (Settings): 'pair' (side-by-side, the
# default), 'stacked', or 'both'. Only the selected layout(s) are rendered — the
# workbook tabs AND the loose PNGs — so "side-by-side only" leaves no stacked
# clutter. Each selected layout becomes one tab PER COMPARISON COLUMN.
LAYOUTS = ("pair", "stacked", "both")
DEFAULT_LAYOUT = "pair"
# img_key -> (embedded display width, human label, per-column-tab suffix used
# ONLY when both layouts are present, so the two tabs for a column stay distinct).
_LAYOUT_SPEC = {"stacked": (_EMBED_W, "stacked", " (stacked)"),
                "pair": (_EMBED_W_PAIR, "side-by-side", " (side-by-side)")}


def normalize_layout(layout):
    """The chosen workbook image layout, defaulting anything unknown to 'pair'."""
    return layout if layout in LAYOUTS else DEFAULT_LAYOUT


def _layout_keys(layout):
    """The img_key(s) to render/embed for `layout`, in reading order (stacked
    before side-by-side when both are selected)."""
    layout = normalize_layout(layout)
    if layout == "stacked":
        return ("stacked",)
    if layout == "both":
        return ("stacked", "pair")
    return ("pair",)


FLAVOR_TSN = "tsn"        # a TSMIS export against the normalized TSN library
FLAVOR_SELF = "self"      # the PDF-vs-Excel self check: both sides are TSMIS
FLAVORS = (FLAVOR_TSN, FLAVOR_SELF)


def self_capable(row_key):
    """Whether the PDF-vs-Excel self check of `row_key` can be illustrated —
    the adapter has to expose the SELF comparator's own loader pair and schema,
    so evidence and comparison can never diverge (CMP-AUD-107)."""
    if not capable(row_key):
        return False
    return hasattr(adapter_for(row_key), "load_sides_self")


def tsmis_source_role(row_key):
    """Which TSMIS export a row's comparison actually READ: 'pdf' or 'excel'.

    CMP-AUD-210: both of a report's rows used to be evidenced from the
    PDF-edition export, and a candidate was dropped whenever that print
    disagreed with the compared value — so anything the Excel export has and the
    print does not (Highway Sequence route 037's Description is the censused
    case) could never be illustrated at all. An Excel row is now evidenced from
    the workbook it was compared from.
    """
    return "pdf" if str(row_key).endswith("_pdf") else "excel"


def capable(row_key):
    return row_key in _ADAPTER_MODULES


def rows():
    return sorted(_ADAPTER_MODULES)


def adapter_for(row_key):
    import importlib
    return importlib.import_module(_ADAPTER_MODULES[row_key])


def pdf_subdir_for(row_key):
    """The TSMIS export subdir whose per-route PDFs illustrate this row."""
    return TSMIS_PDF_SUBDIR[row_key]


def tsn_pdf_dir(row_key):
    report = TSN_PDF_REPORT[row_key]
    if report in _TSN_PDFS_IN_RAW:
        return paths.tsn_library_raw_dir(report)
    return paths.tsn_library_pdf_dir(report)


def clamp_examples(n):
    try:
        return max(MIN_EXAMPLES, min(MAX_EXAMPLES, int(n)))
    except (TypeError, ValueError):
        return DEFAULT_EXAMPLES


def availability():
    """Cheap probe for the GUI toggle: which rows support evidence and, PER
    REPORT, whether that report's TSN prints are in place (the TSMIS side varies
    per run/day and is reported per cell instead). `ready` = deps present and at
    least one report has its prints — the toggle enables then, and a still-empty
    report notes per cell where to drop its prints. The aggregate `tsn_pdfs`/`dir`
    keys stay for the hint's simple cases."""
    reports = []
    total = 0
    for key, label in sorted(_TSN_PDF_LABELS.items()):
        d = (paths.tsn_library_raw_dir(key) if key in _TSN_PDFS_IN_RAW
             else paths.tsn_library_pdf_dir(key))
        try:
            n = sum(1 for _ in Path(d).glob("*.pdf"))
        except OSError:          # silent-ok: a pure probe; unreadable = not ready
            n = 0
        total += n
        reports.append({"key": key, "label": label, "tsn_pdfs": n, "dir": str(d),
                        "source": "raw" if key in _TSN_PDFS_IN_RAW else "pdf"})
    return {"rows": rows(), "tsn_pdfs": total,
            "ready": _DEPS_OK and any(r["tsn_pdfs"] for r in reports),
            "dir": next((r["dir"] for r in reports if not r["tsn_pdfs"]),
                        reports[0]["dir"] if reports else ""),
            "reports": reports, "row_reports": dict(TSN_PDF_REPORT),
            "deps_ok": _DEPS_OK}


def sibling_paths(comparison_path):
    """(workbook, image folder) next to a comparison workbook — the same
    naming family as its '(formulas).xlsx' sibling."""
    p = Path(comparison_path)
    return (p.with_name(f"{p.stem} (evidence){p.suffix}"),
            p.with_name(f"{p.stem} (evidence images)"))


def _pdf_source_files(tsmis_pdf_dir, tsn_dir):
    """The exact PDF paths evidence discovery can consume, in stable order."""
    return (tuple(sorted(Path(tsmis_pdf_dir).glob("*.pdf"))),
            tuple(sorted(Path(tsn_dir).glob("*.pdf"))))


def _ensure_pdf_source_set(tsmis_pdf_dir, tsn_dir, expected):
    """Fail when a PDF is added/removed while the evidence generation runs."""
    current = _pdf_source_files(tsmis_pdf_dir, tsn_dir)
    if artifact_store.canonical_path_identities((*current[0], *current[1])) != expected:
        raise ValueError(
            "Refusing to publish evidence: the discovered PDF source set "
            "changed while evidence was rendering. Re-run after the PDF "
            "folders are stable.")


_DIGEST_CHUNK = 1 << 20


def _pdf_content_digests(paths):
    """{resolved_path: sha256} for the existing PDFs among `paths`."""
    out = {}
    for p in paths:
        p = Path(p)
        try:
            if not p.is_file():
                continue
            h = hashlib.sha256()
            with open(p, "rb") as f:
                for chunk in iter(lambda: f.read(_DIGEST_CHUNK), b""):
                    h.update(chunk)
            out[str(p.resolve())] = h.hexdigest()
        except OSError:              # unreadable now -> treat as changed at check
            out[str(p.resolve())] = None
    return out


class _ReadSet:
    """The private copy of every PDF one evidence generation will read.

    CMP-AUD-098: digesting the live file before parsing and again at the commit
    boundary cannot see an A→B→A swap — digest A, parse B, restore A, and the
    check passes on bytes nobody rendered from. So the bytes are copied ONCE into
    a directory only this run knows about, the COPIES are digested, and every
    later locate/render call reads the copy. Correctness stops depending on what
    the live path does next.

    The commit-boundary check keeps a narrower, honest job: prove the live
    sources still equal what we copied, so the published manifest describes the
    files the user actually has (a source that changed mid-run makes the evidence
    a picture of superseded bytes, which is a refusal, not a render fault).

    Basenames are preserved — the adapters look their PDFs up by name.
    """

    def __init__(self, root, tsmis_dir, tsn_dir, digests, members, identity,
                 fs_identity, extra=()):
        self.root = root
        self.tsmis_dir = tsmis_dir
        self.tsn_dir = tsn_dir
        # {original path -> its snapshot copy} for sources addressed by path
        # rather than discovered in a folder (the compared Excel workbooks).
        self.extra = dict(extra)
        self.digests = digests          # {resolved ORIGINAL path: sha256 of copy}
        self.members = members          # manifest read set, stable order
        self._identity = identity
        self._fs_identity = fs_identity

    def snapshot_of(self, path):
        """The private copy of a by-path source, or the original when it was
        not copied (a caller that asks for one it did not snapshot)."""
        return self.extra.get(str(path), path)

    def ensure_sources_unchanged(self):
        """Refuse when a live source no longer matches the bytes we copied."""
        current = _pdf_content_digests(self.digests.keys())
        for path, digest in self.digests.items():
            if digest is None or current.get(path) != digest:
                raise ValueError(
                    "Refusing to publish evidence: a source PDF's contents "
                    "changed while evidence was rendering (the images would "
                    "illustrate superseded bytes). Re-run after the PDFs are "
                    "stable.")

    def discard(self):
        """Remove the private copy, never a path that replaced it."""
        if not os.path.lexists(self.root):
            return
        try:
            _ensure_captured_current(self._identity)
            if not owned_dir.is_plain_directory_tree(self.root, self._fs_identity):
                raise ValueError("the evidence read-set snapshot was replaced")
        except (ValueError, owned_dir.OwnershipError) as e:
            log.error("evidence: retained replaced read-set snapshot %s (%s: %s)",
                      self.root, type(e).__name__, e)
            return
        shutil.rmtree(self.root, ignore_errors=True)


def _snapshot_read_set(tsmis_paths, tsn_paths, extra=()):
    """Copy the sources this generation will READ into a private snapshot and
    digest the COPIES (CMP-AUD-098). Returns a `_ReadSet`; the caller must
    `discard()` it. A missing source is simply absent from the snapshot — the
    locate step already treats an unreadable route PDF as a miss. `extra` holds
    sources addressed by path rather than found in a folder (the compared Excel
    workbook, CMP-AUD-210)."""
    root = Path(tempfile.mkdtemp(prefix=".tsmis-evidence-readset-"))
    fs_identity = owned_dir.directory_identity(root)
    if fs_identity is None:
        shutil.rmtree(root, ignore_errors=True)
        raise owned_dir.OwnershipError(
            "The evidence read-set snapshot has no stable local identity; it "
            "was not used.")
    identity = artifact_store.capture_source_identities((root,))
    digests, members, extra_map = {}, [], {}
    try:
        for label, sources in (("tsmis", tsmis_paths), ("tsn", tsn_paths),
                               ("compared", extra)):
            side = root / label
            side.mkdir()
            for src in sources:
                src = Path(src)
                if not src.is_file():
                    continue
                copy = side / src.name
                shutil.copyfile(src, copy)
                digest = artifact_store.content_digest(copy)
                digests[str(src.resolve())] = digest
                members.append(evidence_manifest.Member(
                    name=str(src), size=copy.stat().st_size, sha256=digest))
                if label == "compared":
                    extra_map[str(src)] = copy
    except BaseException:
        shutil.rmtree(root, ignore_errors=True)
        raise
    members.sort(key=lambda m: m.name)
    return _ReadSet(root, root / "tsmis", root / "tsn", digests,
                    tuple(members), identity, fs_identity, extra_map)


def _safe_sibling_paths(comparison_path, source_paths, captured_sources=(),
                        source_set_check=None, commit_guard=None):
    """Return evidence siblings only after guarding every derived write path.

    Runtime temp/quarantine/fallback names are allocated unpredictably and are
    guarded when created; only the two deterministic public outputs exist here.
    """
    if source_set_check is not None:
        source_set_check()
    wb_path, img_dir = sibling_paths(comparison_path)
    man_path = evidence_manifest.manifest_path(comparison_path)
    _require_output_guard(commit_guard, wb_path, "evidence workbook selection")
    _require_output_guard(commit_guard, img_dir, "evidence image-folder selection")
    _require_output_guard(commit_guard, man_path, "evidence manifest selection")
    artifact_store.ensure_outputs_do_not_alias_sources(
        (wb_path, img_dir, man_path), source_paths,
        directory_destinations=(img_dir,),
        captured_sources=captured_sources, require_sources_current=True)
    return wb_path, img_dir, man_path


def _ensure_captured_current(captured):
    """Verify an exclusively-created temp/quarantine still names our object."""
    artifact_store.ensure_outputs_do_not_alias_sources(
        (), (), captured_sources=captured, require_sources_current=True)


def _require_output_guard(commit_guard, path=None, action="evidence write", **kwargs):
    """Require the caller's exact output lease at a mutation boundary.

    Target-aware guards accept ``path`` plus the same descendant-binding keyword
    arguments as :meth:`owned_dir.OwnershipLease.guard`.  A legacy zero-argument
    guard remains usable only when no identity binding is requested; silently
    dropping an anchor/directory identity would reopen the replacement race this
    guard exists to close.
    """
    if commit_guard is None:
        return
    try:
        if path is None:
            allowed = bool(commit_guard())
        elif kwargs:
            allowed = bool(commit_guard(Path(path), **kwargs))
        else:
            try:
                allowed = bool(commit_guard(Path(path)))
            except TypeError:
                allowed = bool(commit_guard())
    except Exception as e:                       # noqa: BLE001 - fail closed
        raise owned_dir.OwnershipError(
            f"The evidence output guard failed before the {action}; no output "
            "was intentionally changed.") from e
    if not allowed:
        raise owned_dir.OwnershipError(
            f"The app-owned comparison destination changed before the {action}. "
            "The current path was left untouched; retry the comparison.")


def _unique_dir_sibling(path, tag, keep_suffix=False):
    """Choose an unpredictable, currently absent sibling; never delete a collision.

    `keep_suffix` puts the token before a file's extension so a diverted or
    quarantined workbook still opens by double-click.
    """
    path = Path(path)
    stem, suffix = (path.stem, path.suffix) if keep_suffix else (path.name, "")
    for _ in range(32):
        candidate = path.with_name(f"{stem}.{tag}-{secrets.token_hex(8)}{suffix}")
        if not os.path.lexists(candidate):
            return candidate
    raise FileExistsError(
        f"Could not allocate a collision-free evidence {tag} directory beside {path}.")


def _retire_stale_evidence(wb_path, img_dir, source_paths, captured_sources,
                           source_set_check, commit_guard, man_path=None):
    """Remove a now-stale prior evidence set (workbook + image folder + manifest)
    so it can't survive at its canonical name beside a comparison it no longer
    describes (CMP-AUD-106: this generation produced no artifacts, so any prior
    red evidence belongs to a previous one). Guarded like every other evidence
    mutation and quarantined (atomic move) before delete; a locked/foreign/
    aliased artifact is left in place with a logged note rather than
    force-removed. Returns a short note naming what was retired (or couldn't
    be), or None when nothing existed."""
    retired, failed = [], []
    members = [(wb_path, False, "workbook"), (img_dir, True, "image folder")]
    if man_path is not None:
        members.append((man_path, False, "manifest"))
    for path, is_dir, label in members:
        if not os.path.lexists(path):
            continue
        try:
            if source_set_check is not None:
                source_set_check()
            artifact_store.ensure_outputs_do_not_alias_sources(
                (path,), source_paths,
                directory_destinations=(path,) if is_dir else (),
                captured_sources=captured_sources, require_sources_current=True)
            _require_output_guard(
                commit_guard, path, f"stale evidence {label} retirement")
            quarantine = _unique_dir_sibling(path, "retired")
            os.replace(path, quarantine)          # atomic move out of the way
            if is_dir:
                shutil.rmtree(quarantine, ignore_errors=True)
            else:
                Path(quarantine).unlink(missing_ok=True)
            retired.append(label)
        except (OSError, ValueError, owned_dir.OwnershipError) as e:
            log.warning("evidence: could not retire stale %s %s: %s: %s",
                        label, path, type(e).__name__, e)
            failed.append(label)
    parts = []
    if retired:
        parts.append("retired the now-stale evidence " + " + ".join(retired))
    if failed:
        parts.append("could not remove the stale evidence "
                     + " + ".join(failed) + " (locked open?)")
    return "; ".join(parts) if parts else None


# --------------------------------------------------------------------------- #
# generation
# --------------------------------------------------------------------------- #
def _locate_tsmis_sources(adapter, need_tsmis, tsmis_pdf_dir, events):
    """Locate every needed TSMIS row, one parse per route PDF. Returns
    (tsmis_loc, missing_routes), or None when cancelled.

    CMP-AUD-049 (evidence half): a PDF whose own route claims fail to confirm
    the expected route (the adapter raises pdf_table_lib.RouteIdentityError)
    is EXCLUDED — its examples become misses — never captioned as that route;
    a merely unreadable PDF keeps its separate unreadable path."""
    tsmis_loc, missing_routes = {}, set()
    for ri, (route, keys) in enumerate(sorted(need_tsmis.items()), 1):
        if events.is_cancelled():
            return None
        if ri % 10 == 0:
            events.on_log(f"    …TSMIS PDFs {ri}/{len(need_tsmis)}")
        p = adapter.tsmis_pdf_path(tsmis_pdf_dir, route)
        if not p.is_file():
            missing_routes.add(route)
            continue
        try:
            tsmis_loc[route] = adapter.locate_tsmis(p, keys)
        except RouteIdentityError as e:
            log.warning("evidence: %s excluded: %s", p.name, e)
            events.on_log(f"    ⚠ {e} — excluded from evidence")
            missing_routes.add(route)
        except Exception as e:                            # a corrupt route PDF
            log.warning("evidence: %s unparseable: %s: %s",
                        p.name, type(e).__name__, e)
            missing_routes.add(route)
    return tsmis_loc, missing_routes


def generate(row_key, consolidated, tsn_path, comparison_path, tsmis_pdf_dir,
             events, examples=DEFAULT_EXAMPLES, layout=DEFAULT_LAYOUT,
             commit_guard=None, flavor=FLAVOR_TSN):
    """Generate the evidence set for one finished comparison. Returns a result
    dict {note, rendered, fields_ok, fields_with_diffs, misses, workbook,
    folder} — `note` is the one summary line for the run log. `layout` selects
    which image layout(s) to render (Settings: 'pair'/'stacked'/'both').

    `flavor` picks which comparison is being illustrated: 'tsn' (the default —
    a TSMIS export against the normalized TSN library) or 'self' (the
    PDF-vs-Excel self check, where `tsn_path` is the EXCEL consolidated
    workbook and both sides are TSMIS). CMP-AUD-210: each side is evidenced
    from the source that side was actually read from.

    Raises ValueError for a not-runnable setup (missing deps/PDFs); the caller
    treats any failure as a skipped decoration, never a failed comparison."""
    if not _DEPS_OK:
        raise ValueError(DEPS_MSG)
    if not capable(row_key):
        raise ValueError(f"no visual-evidence support for {row_key}")
    adapter = adapter_for(row_key)
    if flavor not in FLAVORS:
        raise ValueError(f"unknown evidence flavor: {flavor!r}")
    if flavor == FLAVOR_SELF and not self_capable(row_key):
        raise ValueError(f"no PDF-vs-Excel evidence support for {row_key}")
    examples = clamp_examples(examples)
    layout = normalize_layout(layout)
    render_keys = _layout_keys(layout)
    tsmis_pdf_dir = Path(tsmis_pdf_dir)
    # CMP-AUD-210: which SOURCE each side of this comparison was read from. The
    # self check diffs two TSMIS renderings, so side A is the print and side B
    # is the workbook; a vs-TSN comparison always renders side B from the TSN
    # print and takes side A from whichever export the row compared.
    role_a = ("pdf" if flavor == FLAVOR_SELF else tsmis_source_role(row_key))
    role_b = ("excel" if flavor == FLAVOR_SELF else "tsn")
    tsn_dir = tsn_pdf_dir(row_key)
    tsmis_pdf_files, tsn_pdf_files = _pdf_source_files(tsmis_pdf_dir, tsn_dir)
    if role_b != "tsn":
        tsn_pdf_files = ()
        tsn_dir = tsmis_pdf_dir          # nothing else is discovered from it
    source_paths = (consolidated, tsn_path, comparison_path,
                    tsmis_pdf_dir, tsn_dir, *tsmis_pdf_files, *tsn_pdf_files)
    initial_pdf_set = artifact_store.canonical_path_identities(
        (*tsmis_pdf_files, *tsn_pdf_files))

    def source_set_check():
        _ensure_pdf_source_set(tsmis_pdf_dir, tsn_dir, initial_pdf_set)

    captured_sources = artifact_store.capture_source_identities(source_paths)
    wb_path, img_dir, man_path = _safe_sibling_paths(
        comparison_path, source_paths, captured_sources, source_set_check,
        commit_guard=commit_guard)
    if role_a == "pdf" and not tsmis_pdf_files:
        raise ValueError(f"no {adapter.REPORT_LABEL} (PDF) export found in "
                         f"{tsmis_pdf_dir} — run that export first")
    if role_b == "tsn" and not tsn_pdf_files:
        raise ValueError(f"no TSN {adapter.REPORT_LABEL} PDFs in {tsn_dir}")

    seed = int.from_bytes(os.urandom(4), "big")
    rng = random.Random(seed)
    log.info("evidence: %s seed=%08x examples=%d layout=%s tsmis=%s tsn=%s",
             row_key, seed, examples, layout, tsmis_pdf_dir, tsn_dir)
    events.on_log(f"  evidence: sampling up to {examples} example(s) per column "
                  f"(seed {seed:08x})…")

    published, ledger, ledger_digest, fields_with_diffs = (
        evidence_ledger.published_universe(comparison_path, adapter, events))

    def record_no_artifacts(state, note, misses=None):
        """CMP-AUD-106: a generation that publishes no images is still a CURRENT
        state. Retire the prior set so it can't survive looking current, then
        record what this generation found — so a reader, a restart, or a later
        run with the evidence toggle OFF can tell 'nothing to illustrate' apart
        from 'evidence never ran'."""
        retire_note = _retire_stale_evidence(
            wb_path, img_dir, source_paths, captured_sources,
            source_set_check, commit_guard, man_path)
        if retire_note:
            note += f" — {retire_note}"
        manifest_note = _commit_manifest(
            man_path, evidence_manifest.build(
                state=state, report=adapter.REPORT_LABEL,
                comparison_path=comparison_path, ledger_digest=ledger_digest,
                reader_version=ledger.reader_version,
                difference_cells=ledger.difference_cells,
                differing_columns=len(fields_with_diffs),
                seed=f"{seed:08x}", layout=layout, examples=examples,
                note=note),
            source_paths, captured_sources, source_set_check, commit_guard)
        if manifest_note:
            note += f"; {manifest_note}"
        events.on_log("  " + note)
        return {"note": note, "rendered": 0, "fields_ok": 0,
                "fields_with_diffs": len(fields_with_diffs),
                "misses": misses or {}, "workbook": None, "folder": None,
                "manifest_state": state, "ledger": ledger,
                "ledger_digest": ledger_digest}

    if not fields_with_diffs:
        return record_no_artifacts(
            evidence_manifest.STATE_NO_DIFFERENCES,
            "evidence: the published comparison counts no differing columns "
            "to illustrate")

    load = (adapter.load_sides_self if flavor == FLAVOR_SELF
            else adapter.load_sides)
    rows_a, rows_b, sidecar, note = load(consolidated, tsn_path)
    if sidecar is None:
        raise ValueError(note or "the TSN workbook carries no district info")
    if events.is_cancelled():
        return _cancelled()
    # The adapter PROPOSES rows that can be photographed; every one of them is
    # then checked against the published cell it claims to illustrate. The self
    # check judges cells with the SELF comparator's own schema, which the
    # adapter puts in the sidecar (CMP-AUD-107: never a second equality engine).
    proposed = adapter.enumerate_diffs(
        rows_a, rows_b, sidecar,
        **({"schema": sidecar["schema"]} if flavor == FLAVOR_SELF else {}))
    renderable, rejected = evidence_ledger.reconcile(published, proposed)
    refused = sum(sum(why.values()) for why in rejected.values())
    if refused:
        events.on_log(f"  evidence: {refused:,} proposed example(s) did not "
                      "match the published cell and were not rendered")

    # pick candidates, then group the lookups by source file so each PDF is
    # parsed exactly once
    cand = {}
    misses = {}
    for f in fields_with_diffs:
        pool = renderable.get(f) or []
        if not pool:
            why = evidence_ledger.unrenderable_reason(ledger.for_field(f), rejected.get(f))
            if why:
                misses[f] = why
                log.info("evidence: %s — %s", f, why)
            continue
        cand[f] = rng.sample(pool, min(len(pool),
                                       max(examples * 4, examples + 6)))
    if not cand:
        # Every differing column is unrenderable (CMP-AUD-108's duplicate-only
        # shape, or the published cells refused every proposal). Report the
        # published differences and their reasons WITHOUT parsing a single PDF
        # — there is nothing left to locate.
        return record_no_artifacts(
            evidence_manifest.STATE_NO_EXAMPLES,
            f"evidence: {ledger.difference_cells:,} published difference(s) "
            f"across {len(fields_with_diffs)} column(s), none of them with a "
            "row that can be illustrated on its own (reasons in the log)",
            misses)
    need_tsmis = {}
    need_tsn_routes, need_tsn_keys = {}, {}
    for f in cand:
        for ex in cand[f]:
            need_tsmis.setdefault(ex["route"], set()).add(ex["key"])
            need_tsn_routes.setdefault(ex["dist"], set()).add(ex["route"])
            need_tsn_keys.setdefault(ex["dist"], set()).add(
                (ex["cnty"], ex["route"], ex["key"]))

    # CMP-AUD-098/112: copy every source this generation will READ into a
    # private snapshot NOW and digest the COPIES. Every locate and render below
    # reads that snapshot, so what the images illustrate cannot be changed by
    # anything that happens to the live files afterwards. CMP-AUD-210: only the
    # sources the compared SIDES actually came from are copied — a companion
    # print that took no part in this comparison is not a source of it.
    excel_books = ([consolidated] if role_a == "excel" else []) \
        + ([tsn_path] if role_b == "excel" else [])
    read_set = _snapshot_read_set(
        ([adapter.tsmis_pdf_path(tsmis_pdf_dir, r) for r in need_tsmis]
         if role_a == "pdf" else []),
        tsn_pdf_files, extra=excel_books)
    try:
        tsmis_loc, missing_routes = {}, set()
        excel_a = excel_b = ({}, [])
        if role_a == "excel":
            # Address the compared cells in the workbook the comparison read,
            # not in a companion print that may hold something else — or
            # nothing — at that row.
            excel_a = _excel_rows_at(
                read_set.snapshot_of(consolidated),
                {e["row_index"] for f in cand for e in cand[f]
                 if e.get("row_index") is not None})
            events.on_log(f"  evidence: addressing {len(excel_a[0])} row(s) in "
                          "the compared TSMIS workbook…")
        else:
            events.on_log(f"  evidence: locating candidates in {len(need_tsmis)} "
                          "TSMIS PDF(s)…")
            located = _locate_tsmis_sources(adapter, need_tsmis,
                                            read_set.tsmis_dir, events)
            if located is None:
                return _cancelled()
            tsmis_loc, missing_routes = located
        if missing_routes:
            events.on_log(f"    note: no readable/confirmable TSMIS PDF for route(s) "
                          f"{', '.join(sorted(missing_routes))} — sampling around them")
        if role_b == "excel":
            excel_b = _excel_rows_at(
                read_set.snapshot_of(tsn_path),
                {e["row_index_b"] for f in cand for e in cand[f]
                 if e.get("row_index_b") is not None})
            events.on_log(f"  evidence: addressing {len(excel_b[0])} row(s) in "
                          "the compared TSMIS Excel workbook…")
            dist_index, tsn_loc = {}, {}
            need_tsn_keys = {}
        else:
            events.on_log(f"  evidence: reading {len(need_tsn_keys)} TSN "
                          "district print(s)…")
        dist_index = ({} if role_b == "excel"
                      else adapter.district_index(read_set.tsn_dir, events))
        tsn_loc = {}
        # The district prints are the slow half (word extraction on every page), so
        # narrate each one — a stalled run must name where it stalled.
        for di, dist in enumerate(sorted(need_tsn_keys), 1):
            if events.is_cancelled():
                return _cancelled()
            p = dist_index.get(dist)
            if p is None:
                continue
            try:
                tsn_loc[dist] = adapter.locate_tsn(p, need_tsn_routes[dist],
                                                   need_tsn_keys[dist])
                events.on_log(f"    …TSN district {dist}: "
                              f"{sum(len(v) for v in tsn_loc[dist].values())} "
                              f"candidate row(s) ({di}/{len(need_tsn_keys)})")
            except Exception as e:
                log.warning("evidence: %s unparseable: %s: %s",
                            p.name, type(e).__name__, e)

        # render into a temp folder; swap in only on success (keep-last-good)
        source_set_check()
        _require_output_guard(commit_guard, img_dir.parent,
                              "evidence output-folder creation")
        img_dir.parent.mkdir(parents=True, exist_ok=True)
        _require_output_guard(commit_guard, img_dir.parent,
                              "evidence output-folder creation")
        tmp_dir = Path(tempfile.mkdtemp(
            prefix=f".{img_dir.name}.tmp-", dir=img_dir.parent))
        tmp_dir_fs_identity = owned_dir.directory_identity(tmp_dir)
        if tmp_dir_fs_identity is None:
            raise owned_dir.OwnershipError(
                "The temporary evidence image folder has no stable local identity; "
                "it was not used.")
        tmp_dir_identity = artifact_store.capture_source_identities((tmp_dir,))
        _require_output_guard(
            commit_guard, tmp_dir, "temporary evidence-folder creation",
            directory_identity=tmp_dir_fs_identity)
        page_cache = {}
        entries = []
        rendered = 0
        try:
            for fi, f in enumerate(cand, 1):
                if events.is_cancelled():
                    return _cancelled()
                got, reasons = 0, []
                for ex in cand[f]:
                    if got >= examples:
                        break
                    _require_output_guard(
                        commit_guard, tmp_dir, "evidence image write",
                        directory_identity=tmp_dir_fs_identity)
                    ok, reason = _try_example(
                        adapter, ex, f, tsmis_loc, tsn_loc, dist_index,
                        read_set.tsmis_dir, tmp_dir, got + 1, page_cache,
                        render_keys, commit_guard=commit_guard,
                        out_dir_identity=tmp_dir_fs_identity,
                        tsmis_role=role_a, excel_rows=excel_a[0],
                        excel_header=excel_a[1],
                        consolidated_name=Path(consolidated).name,
                        role_b=role_b, excel_b=excel_b[0],
                        excel_b_header=excel_b[1],
                        excel_b_name=Path(tsn_path).name)
                    if ok:
                        got += 1
                        rendered += 1
                        entries.append(ok)
                    else:
                        reasons.append(reason)
                if not got:
                    misses[f] = _summarize_reasons(reasons)
                    log.info("evidence: %s — no verifiable example (%s)",
                             f, misses[f])
                if fi % 8 == 0:
                    events.on_log(f"  evidence: {fi}/{len(cand)} columns done…")
            if events.is_cancelled():
                return _cancelled()
            if not rendered:
                # CMP-AUD-108: the published differences are still REPORTED — the
                # columns and their reasons, never a silent zero. CMP-AUD-106: and
                # the prior set is retired, with the state recorded.
                return record_no_artifacts(
                    evidence_manifest.STATE_NO_EXAMPLES,
                    f"evidence: {ledger.difference_cells:,} published "
                    f"difference(s) across {len(fields_with_diffs)} column(s), "
                    "none of them with a verifiable example to render "
                    "(reasons in the log)", misses)
            # CMP-AUD-109 + 112: recheck at the commit boundary — cancellation and
            # the source PDFs — immediately before publishing the SET. A late
            # cancel leaves the previous evidence untouched (keep-last-good); a
            # source that no longer matches the snapshot the images were rendered
            # from aborts, so evidence never illustrates superseded bytes.
            if events.is_cancelled():
                return _cancelled()
            read_set.ensure_sources_unchanged()
            # CMP-AUD-208: every published item names the two SOURCE rows it was
            # taken from, resolved through the comparison's own opaque row token.
            evidence_ledger.attach_source_rows(published, entries)
            sampled = collections.Counter(e["field"] for e in entries)

            def manifest_for(published_wb, published_images):
                """Measure what actually landed at the canonical names — the
                manifest vouches for the committed bytes, not for the staged
                ones it hoped to commit."""
                return evidence_manifest.build(
                    state=evidence_manifest.STATE_RENDERED,
                    report=adapter.REPORT_LABEL,
                    comparison_path=comparison_path,
                    ledger_digest=ledger_digest,
                    reader_version=ledger.reader_version,
                    difference_cells=ledger.difference_cells,
                    differing_columns=len(fields_with_diffs),
                    read_set=read_set.members, workbook=published_wb,
                    images=tuple(sorted(
                        (evidence_manifest.member_for(p)
                         for p in Path(published_images).iterdir() if p.is_file()),
                        key=lambda m: m.name)),
                    seed=f"{seed:08x}", layout=layout, examples=examples)

            publication = _publish_evidence_set(
                wb_path, img_dir, tmp_dir, entries, misses, dict(
                    comparison=Path(comparison_path).name,
                    report=adapter.REPORT_LABEL, seed=f"{seed:08x}",
                    examples=examples, tsmis_dir=str(tsmis_pdf_dir),
                    tsn_dir=str(tsn_dir),
                    sides=published.side_labels,
                    reader_version=ledger.reader_version,
                    ledger_digest=ledger_digest,
                    ledger=evidence_ledger.ledger_rows(ledger, adapter.FIELDS, sampled),
                    ledger_totals=ledger),
                layout, source_paths, captured_sources, source_set_check,
                commit_guard, tmp_dir_fs_identity, man_path, manifest_for)
        finally:
            if os.path.lexists(tmp_dir):
                try:
                    _ensure_captured_current(tmp_dir_identity)
                    artifact_store.ensure_outputs_do_not_alias_sources(
                        (tmp_dir,), source_paths, directory_destinations=(tmp_dir,),
                        captured_sources=captured_sources,
                        require_sources_current=True)
                    _require_output_guard(
                        commit_guard, tmp_dir, "temporary evidence-folder cleanup",
                        directory_identity=tmp_dir_fs_identity)
                except (ValueError, owned_dir.OwnershipError):
                    log.error("evidence: retained replaced/unsafe temp path instead of "
                              "deleting foreign or source content: %s", tmp_dir)
                else:
                    shutil.rmtree(tmp_dir, ignore_errors=True)

        fields_ok = len({e["field"] for e in entries})
        promoted = publication["status"] == "promoted"
        note = (f"evidence: {rendered} example(s) across {fields_ok}/"
                f"{len(fields_with_diffs)} differing column(s)"
                + (f" → {wb_path.name}" if promoted else ""))
        if misses:
            note += (f" — {len(misses)} column(s) had no verifiable example "
                     "(reasons in the workbook)")
        if publication.get("note"):
            note += f"; {publication['note']}"
        events.on_log("  " + note)
        # CMP-AUD-109: the returned workbook/folder are the ACTUAL committed
        # canonical paths (None when the set diverted to .new), not an unconditional
        # success claim; `status` names promoted vs diverted.
        return {"note": note, "rendered": rendered, "fields_ok": fields_ok,
                "fields_with_diffs": len(fields_with_diffs), "misses": misses,
                "workbook": publication["workbook"], "folder": publication["folder"],
                "status": publication["status"],
                "manifest_state": (evidence_manifest.STATE_RENDERED if promoted
                                   else evidence_manifest.STATE_DIVERTED),
                "ledger": ledger, "ledger_digest": ledger_digest}
    finally:
        read_set.discard()




def _cancelled():
    return {"note": "evidence: cancelled — previous evidence files left as-is",
            "rendered": 0, "fields_ok": 0, "fields_with_diffs": 0,
            "misses": {}, "workbook": None, "folder": None}




def _summarize_reasons(reasons):
    if not reasons:
        return "no candidates"
    uniq = []
    for r in reasons:
        if r not in uniq:
            uniq.append(r)
    return "; ".join(uniq[:3])


_QUOTE_TOKEN_RE = re.compile(r"''|[\"']")
_QUOTE_NAMES = {"''": "'' (two apostrophes)", '"': '" (a quotation mark)',
                "'": "' (one apostrophe)"}


def _quote_note(va, vb, b_label="TSN"):
    """The one invisible diff class: two values that differ ONLY in quote
    characters ('' vs " vs ') print near-identically, so a verified-real
    difference reads as a false positive (the censused case: Intersection
    Detail KER 046 @ 50.904 — TSMIS ''F'' ST vs TSN "F" ST, the single
    statewide instance). Returns the header line naming both sides'
    characters, or "" for every other pair."""
    a, b = str(va or ""), str(vb or "")
    if a == b or _QUOTE_TOKEN_RE.sub("'", a) != _QUOTE_TOKEN_RE.sub("'", b):
        return ""
    pair = next(((x, y) for x, y in zip(_QUOTE_TOKEN_RE.findall(a),
                                        _QUOTE_TOKEN_RE.findall(b)) if x != y),
                None)
    if pair is None:                    # unreachable when a != b; kept safe
        return ""
    return ("REAL difference, in the quote characters only: TSMIS prints "
            f"{_QUOTE_NAMES[pair[0]]} where {b_label} prints "
            f"{_QUOTE_NAMES[pair[1]]}")


def _tsmis_excel_side(adapter, ex, field, excel_rows, excel_header,
                      consolidated_name, index_key="row_index",
                      value_key="va", side_label="TSMIS (Excel)"):
    """Render one side from the WORKBOOK the comparison read (CMP-AUD-210).

    Returns (image, label, address, None) or (None, None, None, reason). The
    cell is only rendered once its own workbook value, put through the adapter's
    projection, equals the value the comparison compared — the Excel counterpart
    of the PDF side's parse-back check. A companion print's disagreement is
    irrelevant here, which is the point: an Excel value the print never carried
    stays evidenceable as Excel truth.
    """
    index = ex.get(index_key)
    located = excel_rows.get(index) if index is not None else None
    if located is None:
        return None, None, None, "row not found in the compared TSMIS workbook"
    sheet, excel_row, values = located
    # The compared column is resolved through the COMPARATOR'S OWN column
    # resolution when the adapter exposes it (excel_column_for — the same
    # position map its loader reads the workbook with, under the loader's own
    # exact-header gate), because the compared labels are the COMPARISON's
    # shared names while the workbook carries the SITE's labels — Intersection
    # Detail compares 'PM'/'PR'/'HG' where the workbook says 'Post Mile'/'PP'/
    # 'H/G', so the old exact-label lookup silently dropped those columns'
    # Excel-side evidence, and Ramp Detail's label-shifted header could box a
    # NEIGHBOURING blank cell. Adapters without the hook keep the exact-label
    # default; a None from the hook is the honest refusal (a derived column
    # with no single workbook cell, or a workbook the comparator would refuse).
    resolve = getattr(adapter, "excel_column_for", None)
    if resolve is not None:
        col = resolve(field, excel_header)
        if col is None:
            return None, None, None, (
                "the compared column has no single cell in this workbook "
                "(a derived column, or a header the comparison would refuse)")
    elif field in excel_header:
        col = excel_header.index(field)
    else:
        return None, None, None, "the compared column is not in the workbook header"
    if col >= len(values):
        return None, None, None, "the workbook row is short of the compared column"
    if adapter.project(field, values[col]) != adapter.project(field, ex[value_key]):
        return None, None, None, ("the compared TSMIS workbook cell no longer "
                                  "holds the compared value")
    key_cols = [0]
    if resolve is not None:
        key_col = resolve(adapter.KEY_LABEL, excel_header)
        if key_col is not None:
            key_cols.append(key_col)
    elif adapter.KEY_LABEL in excel_header:
        key_cols.append(excel_header.index(adapter.KEY_LABEL))
    img = _excel_strip(excel_header, values, col, key_cols)
    address = f"{sheet}!{_column_letter(col + 1)}{excel_row}"
    label = f"{side_label}  —  {consolidated_name} · {address}"
    return img, label, address, None


def _try_example(adapter, ex, field, tsmis_loc, tsn_loc, dist_index,
                 tsmis_pdf_dir, out_dir, k, page_cache,
                 render_keys=("stacked", "pair"), commit_guard=None,
                 out_dir_identity=None, tsmis_role="pdf", excel_rows=None,
                 excel_header=(), consolidated_name="", role_b="tsn",
                 excel_b=None, excel_b_header=(), excel_b_name=""):
    """Verify one candidate end-to-end and render the selected layout(s).
    `render_keys` is the subset of ('stacked', 'pair') the user chose.
    `tsmis_role`/`role_b` say which source each compared SIDE was read from, so
    each side is evidenced from THAT source (CMP-AUD-210). Returns
    (entry_dict, None) on success, (None, reason) otherwise."""
    dist = cnty = ""
    n_pdf = None
    if role_b == "excel":
        n_img, n_label, b_address, reason = _tsmis_excel_side(
            adapter, ex, field, excel_b or {}, excel_b_header, excel_b_name,
            index_key="row_index_b", value_key="vb",
            side_label="TSMIS (Excel)")
        if reason is not None:
            return None, reason.replace("TSMIS workbook",
                                        "TSMIS Excel workbook")
        n_verified = f"the Excel cell {b_address}"
    else:
        nrecs = tsn_loc.get(ex["dist"], {}).get(
            (ex["cnty"], ex["route"], ex["key"]), [])
        if len(nrecs) != 1:
            return None, "row not found uniquely in the TSN district print"
        nrec = nrecs[0]
        nv = adapter.tsn_value(nrec, field)
        if nv != ex["vb"]:
            return None, "the TSN print differs from the TSN workbook at this cell"
        nb = adapter.tsn_box(nrec, field)
        if nb is None:
            return None, "the TSN record's geometry isn't evidence-grade here"
        npage, nbox, nyspan, nxspan = nb
        # A record that names its own source print (the Highway Log's per-print
        # routing) wins over the district index; likewise its district/county
        # provenance (learned from the print's own headers) enriches the captions.
        n_pdf = Path(nrec.get("src") or dist_index[ex["dist"]])
        dist = nrec.get("dist") or ex["dist"]
        cnty = nrec.get("cnty") or ex["cnty"]
        n_img = _strip(n_pdf, npage, nbox, nyspan, nxspan, page_cache)
        n_label = (f"TSN  —  {n_pdf.name} · page {npage} · "
                   f"{(cnty + '-') if cnty else ''}{ex['route']}")
        n_verified = "the TSN print"

    if tsmis_role == "excel":
        t_img, t_label, address, reason = _tsmis_excel_side(
            adapter, ex, field, excel_rows or {}, excel_header,
            consolidated_name)
        if reason is not None:
            return None, reason
        t_verified = f"the workbook cell {address}"
    else:
        trecs = tsmis_loc.get(ex["route"], {}).get(ex["key"], [])
        if len(trecs) != 1:
            return None, ("no readable TSMIS PDF for the route" if
                          ex["route"] not in tsmis_loc
                          else "row not found uniquely in the TSMIS PDF")
        trec = trecs[0]
        tb = adapter.tsmis_box(trec, field)
        if tb is None:
            return None, "record on an approximate-geometry page"
        tv = adapter.tsmis_value(trec, field)
        if tv != ex["va"]:
            return None, ("the TSMIS PDF prints a different value than the "
                          "compared export")
        tpage, tbox, tyspan, txspan = tb
        t_pdf = adapter.tsmis_pdf_path(tsmis_pdf_dir, ex["route"])
        t_img = _strip(t_pdf, tpage, tbox, tyspan, txspan, page_cache)
        t_label = f"TSMIS (PDF)  —  {t_pdf.name} · page {tpage}"
        t_verified = "the TSMIS print"

    b_name = "TSMIS (Excel)" if role_b == "excel" else "TSN"
    title = (f"{field} — TSMIS '{ex['va'] or '(blank)'}'  vs  "
             f"{b_name} '{ex['vb'] or '(blank)'}'")
    note = _quote_note(ex["va"], ex["vb"], b_name)
    where = f" — TSN district D{dist} ({cnty})" if dist or cnty else ""
    sub = (f"Route {ex['route']} @ {ex['key']} — {t_verified} and "
           f"{n_verified} re-read and verified against the compared "
           f"values{where}")
    safe = re.sub(r"[^A-Za-z0-9]+", "_", field).strip("_")
    guard_kwargs = ({"anchor_path": out_dir,
                     "anchor_identity": out_dir_identity}
                    if out_dir_identity is not None else {})
    entry = {"field": field, "route": ex["route"], "key": ex["key"],
             "va": ex["va"], "vb": ex["vb"], "note": note}
    # CMP-AUD-208: carry the published coordinates this image illustrates, so
    # the workbook can name the exact cell instead of just the values.
    for name in ("published_row", "published_occurrence", "published_state",
                 "published_token"):
        if name in ex:
            entry[name] = ex[name]
    # Render ONLY the layout(s) the user chose (Settings). The costly page
    # strips above are shared; only the compose+save differ per layout.
    if "stacked" in render_keys:
        stacked = out_dir / f"{safe}_{k}_stacked.png"
        _require_output_guard(commit_guard, stacked,
                              "stacked evidence-image write", **guard_kwargs)
        _compose_stacked(title, sub, t_label, t_img, n_label, n_img, stacked,
                         note=note)
        _require_output_guard(commit_guard, stacked,
                              "evidence image verification", **guard_kwargs)
        entry["stacked"] = stacked.name
    if "pair" in render_keys:
        pair = out_dir / f"{safe}_{k}_pair.png"
        _require_output_guard(commit_guard, pair,
                              "paired evidence-image write", **guard_kwargs)
        _compose_pair(title, sub, t_label, t_img, n_label, n_img, pair,
                      note=note)
        _require_output_guard(commit_guard, pair,
                              "evidence image verification", **guard_kwargs)
        entry["pair"] = pair.name
    return entry, None


# --------------------------------------------------------------------------- #
# rendering
# --------------------------------------------------------------------------- #
def _render_page(path, page_no, cache):
    key = (str(path), page_no)
    if key not in cache:
        if len(cache) >= _PAGE_CACHE_MAX:
            cache.pop(next(iter(cache)))
        with pdfplumber.open(path) as pdf:
            cache[key] = pdf.pages[page_no - 1].to_image(
                resolution=_RES).original.convert("RGB")
    return cache[key]


def _crop_window(img_w, img_h, cell_box, record_yspan):
    """The strip's crop box (pixels): a FULL-WIDTH page band around the record.
    The adapters' xspan covers only the record's own words, so cropping to it
    clipped whatever printed beyond them — a blank cell's red box (drawn where
    the value WOULD print) and the neighbors' longer text both fell outside.
    The band must span the page and stretch vertically over the cell box."""
    _x0, y0, _x1, y1 = cell_box
    ry0, ry1 = record_yspan
    top = min(ry0 - _CTX_PT, y0 - 4)
    bottom = max(ry1 + _CTX_PT + 2, y1 + 4)
    return (0, int(max(0, top * _SC)), img_w, int(min(img_h, bottom * _SC)))


def _strip(path, page_no, cell_box, record_yspan, xspan, cache):
    """A context crop around the record: gray box = the record's printed
    line(s) (its word extent, `xspan`), red box = the compared cell."""
    img = _render_page(path, page_no, cache).copy()
    d = ImageDraw.Draw(img)
    rx0, rx1 = xspan
    ry0, ry1 = record_yspan
    d.rectangle([rx0 * _SC, (ry0 - 1.5) * _SC, rx1 * _SC, (ry1 + 1.5) * _SC],
                outline=(150, 150, 150), width=2)
    x0, y0, x1, y1 = cell_box
    d.rectangle([x0 * _SC, y0 * _SC, x1 * _SC, y1 * _SC],
                outline=(220, 20, 20), width=4)
    return img.crop(_crop_window(img.width, img.height, cell_box, record_yspan))


def _excel_rows_at(workbook_path, wanted):
    """{data_index: (sheet, excel_row_no, [values])} for the wanted 0-based DATA
    row indexes of a consolidated export (CMP-AUD-210).

    The export is a flat labelled table — header on row 1, one data row after —
    so the loader's row order IS the sheet's row order. Nothing downstream
    trusts that: the caller renders a cell only once its own value matches what
    the comparison compared, which is the Excel counterpart of the PDF side's
    parse-back check.
    """
    if not wanted:
        return {}, []
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header, out = [], {}
        last = max(wanted)
        for excel_row, row in enumerate(
                ws.iter_rows(min_row=1, max_row=last + 2, values_only=True), 1):
            if excel_row == 1:
                header = ["" if v is None else str(v) for v in row]
                continue
            index = excel_row - 2
            if index in wanted:
                out[index] = (ws.title, excel_row,
                              ["" if v is None else str(v) for v in row])
        return out, header
    finally:
        wb.close()


def _column_letter(index):
    """1-based column index -> its Excel letter, so the address is one a user
    can type into the Name Box."""
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# The cell strip's geometry, in pixels.
_XL_PAD = 16
_XL_ROW_H = 46
_XL_HEAD_H = 34
_XL_MIN_COL_W = 92
_XL_CHAR_W = 13                # rough per-character width at the value font size
_XL_CONTEXT_COLS = 3           # columns shown either side of the compared one


def _excel_strip(header, values, target_index, key_indexes=(0,)):
    """Render one CONSOLIDATED-workbook row as a boxed cell strip.

    The Excel counterpart of `_strip`: same conventions, so the two sides of a
    composed image read alike — gray box around the row that was compared, red
    box around the compared cell. It is a rendering of the workbook's own
    values, never a screenshot of Excel, and the caller's label says so.
    """
    show = sorted(set(key_indexes) | {
        i for i in range(max(0, target_index - _XL_CONTEXT_COLS),
                         min(len(values), target_index + _XL_CONTEXT_COLS + 1))})
    widths = [max(_XL_MIN_COL_W,
                  _XL_CHAR_W * max(len(str(header[i] if i < len(header) else "")),
                                   len(str(values[i])), 1) + 18)
              for i in show]
    w = sum(widths) + 2 * _XL_PAD
    h = _XL_HEAD_H + _XL_ROW_H + 2 * _XL_PAD
    img = Image.new("RGB", (w, h), (255, 255, 255))
    d = ImageDraw.Draw(img)
    head_font, cell_font = _font(15, True), _font(19)
    x = _XL_PAD
    gap_after = None
    for i, col in enumerate(show):
        cw = widths[i]
        top = _XL_PAD
        d.rectangle([x, top, x + cw, top + _XL_HEAD_H], fill=(242, 242, 242),
                    outline=(196, 196, 196))
        label = str(header[col] if col < len(header) else "")
        d.text((x + 8, top + 8), label[:24], font=head_font, fill=(70, 70, 70))
        d.rectangle([x, top + _XL_HEAD_H, x + cw, top + _XL_HEAD_H + _XL_ROW_H],
                    outline=(196, 196, 196))
        text = str(values[col])
        d.text((x + 8, top + _XL_HEAD_H + 12), text[:26], font=cell_font,
               fill=(20, 20, 20))
        if col == target_index:
            d.rectangle([x, top + _XL_HEAD_H, x + cw,
                         top + _XL_HEAD_H + _XL_ROW_H],
                        outline=(220, 20, 20), width=4)
        # A break in the shown columns is marked, so the strip never implies the
        # key column sits next to the compared one when it does not.
        if gap_after is not None and col != gap_after + 1:
            d.line([x - 1, top, x - 1, top + _XL_HEAD_H + _XL_ROW_H],
                   fill=(120, 120, 120), width=3)
        gap_after = col
        x += cw
    d.rectangle([_XL_PAD, _XL_PAD + _XL_HEAD_H, w - _XL_PAD,
                 _XL_PAD + _XL_HEAD_H + _XL_ROW_H],
                outline=(150, 150, 150), width=2)
    return img


_FONT_WARNED = False


def _font(size, bold=False):
    global _FONT_WARNED
    name = "arialbd.ttf" if bold else "arial.ttf"
    windir = os.environ.get("WINDIR", r"C:\Windows")
    try:
        return ImageFont.truetype(str(Path(windir) / "Fonts" / name), size)
    except OSError:
        if not _FONT_WARNED:
            _FONT_WARNED = True
            log.info("evidence: Arial not found; using the built-in font")
        return ImageFont.load_default()


def _scaled(im, width):
    if im.width <= width:
        return im
    return im.resize((width, round(im.height * width / im.width)), Image.LANCZOS)


_NOTE_H = 30            # extra header height when the quote-characters note line shows


def _header(canvas, w, title, sub, note=""):
    """The title block; returns the y where content starts. `note` (the
    _quote_note line) renders dark red under the subtitle — it flags a REAL
    difference the printed values themselves don't show."""
    d = ImageDraw.Draw(canvas)
    d.text((16, 12), title, font=_font(26, True), fill=(20, 20, 20))
    d.text((16, 50), sub, font=_font(17), fill=(90, 90, 90))
    if note:
        d.text((16, 80), note, font=_font(17, True), fill=(165, 20, 20))
    return 84 + (_NOTE_H if note else 0)


def _compose_stacked(title, sub, top_label, top_img, bot_label, bot_img, out,
                     note=""):
    top_img, bot_img = _scaled(top_img, _STACK_W), _scaled(bot_img, _STACK_W)
    w = max(top_img.width, bot_img.width) + 32
    lab = 30
    hd = 84 + (_NOTE_H if note else 0)
    h = hd + lab + top_img.height + 14 + lab + bot_img.height + 16
    canvas = Image.new("RGB", (w, h), (255, 255, 255))
    y = _header(canvas, w, title, sub, note)
    d = ImageDraw.Draw(canvas)
    for label, im in ((top_label, top_img), (bot_label, bot_img)):
        d.text((16, y + 4), label, font=_font(16, True), fill=(31, 56, 100))
        y += lab
        canvas.paste(im, (16, y))
        d.rectangle([15, y - 1, 16 + im.width, y + im.height],
                    outline=(200, 200, 200), width=1)
        y += im.height + 14
    canvas.save(out)


def _compose_pair(title, sub, l_label, l_img, r_label, r_img, out, note=""):
    l_img, r_img = _scaled(l_img, _PAIR_SIDE_W), _scaled(r_img, _PAIR_SIDE_W)
    lab = 30
    col_h = max(l_img.height, r_img.height)
    w = l_img.width + r_img.width + 48
    h = 84 + (_NOTE_H if note else 0) + lab + col_h + 16
    canvas = Image.new("RGB", (w, h), (255, 255, 255))
    y0 = _header(canvas, w, title, sub, note)
    d = ImageDraw.Draw(canvas)
    x = 16
    for label, im in ((l_label, l_img), (r_label, r_img)):
        d.text((x, y0 + 4), label, font=_font(16, True), fill=(31, 56, 100))
        canvas.paste(im, (x, y0 + lab))
        d.rectangle([x - 1, y0 + lab - 1, x + im.width, y0 + lab + im.height],
                    outline=(200, 200, 200), width=1)
        x += im.width + 16
    canvas.save(out)


# --------------------------------------------------------------------------- #
# outputs (keep-last-good)
# --------------------------------------------------------------------------- #
def _safe_cell(ws, row, column, value, font=None):
    """Write workbook content that may contain source/user-controlled text."""
    cell = set_safe_literal_cell(ws.cell(row=row, column=column), value)
    if font is not None:
        cell.font = font
    return cell


_ILLEGAL_SHEET_RE = re.compile(r"[\[\]:*?/\\]")


def _sheet_name(base, used, suffix=""):
    """A legal (<=31 chars; no []:*?/\\), unique Excel sheet name for `base`,
    reserving room for `suffix` (the layout tag, present only when both layouts
    show). Collisions after truncation take a numeric disambiguator."""
    base = _ILLEGAL_SHEET_RE.sub("-", str(base)).strip() or "Column"
    room = max(1, 31 - len(suffix))
    name = base[:room] + suffix
    if name not in used:
        used.add(name)
        return name
    for i in range(2, 1000):
        tag = f" ({i}){suffix}"
        cand = base[:max(1, 31 - len(tag))] + tag
        if cand not in used:
            used.add(cand)
            return cand
    cand = (base[:20] + secrets.token_hex(4))[:31]
    used.add(cand)
    return cand


def _column_image_sheets(wb, entries, img_dir, img_key, embed_w, label,
                         fonts, used, suffix=""):
    """Per-COMPARISON-COLUMN tabs of ONE layout: each differing column gets its
    own sheet (named for the field, tagged with the layout `suffix` only when
    both layouts show), holding that column's captioned example images scaled
    to `embed_w`. Columns with no rendered example in this layout are skipped —
    the Summary lists those as misses. Returns [(field, sheet_name)]."""
    bold, small = fonts
    by_field = {}
    for e in entries:
        if e.get(img_key):
            by_field.setdefault(e["field"], []).append(e)
    made = []
    for field, group in by_field.items():
        name = _sheet_name(field, used, suffix)
        made.append((field, name))
        ev = wb.create_sheet(name)
        ev.sheet_properties.tabColor = "C00000"
        _safe_cell(ev, 1, 1,
                   f"{field} — {label} evidence   ·   {len(group)} example(s)",
                   bold)
        _safe_cell(ev, 2, 1,
                   "Red box = the compared cell in each source PDF; gray box = "
                   "the record (its printed lines). Values shown are the "
                   "compared (normalized) forms.", small)
        r = 4
        for e in group:
            note = e.get("note") or ""
            _safe_cell(ev, r, 1, (
                f"route {e['route']} @ {e['key']}   —   "
                f"TSMIS '{e['va']}' vs TSN '{e['vb']}'"
                + (f"   —   {note}" if note else "")), bold)
            img = XLImage(str(img_dir / e[img_key]))
            scale = min(1.0, embed_w / img.width)
            img.width, img.height = int(img.width * scale), int(img.height * scale)
            ev.add_image(img, f"A{r + 1}")
            r += 1 + max(1, round(img.height / _PX_PER_ROW)) + 2
    return made


def _published_cell_text(entry):
    """"Comparison!<row> · occurrence N · state D" for one rendered item."""
    row = entry.get("published_row")
    if not row:
        return ""
    return (f"Comparison!{row} · occurrence "
            f"{entry.get('published_occurrence', 1)} · state "
            f"{entry.get('published_state', '')}")


def _source_rows_text(entry):
    pairs = entry.get("source_rows") or ()
    return "  ·  ".join(f"{side}!{row}" for side, row in pairs if row)


_LEDGER_HEADERS = ("Column", "Counted differences", "With a unique row",
                   "Inside repeated-key groups", "Context cells",
                   "Identical cells", "One-sided cells", "Examples rendered",
                   "Why no example")


def _write_ledger_sheet(wb, info, misses, fonts):
    """The EXHAUSTIVE published accounting (CMP-AUD-209).

    Written from the ledger that was built and hash-bound BEFORE any sample was
    drawn, so the images are a presentation of this record and never a
    substitute for it. Absent when the caller supplied no ledger (an older
    caller); the sheet is the completeness record, not a decoration.
    """
    rows = info.get("ledger")
    totals = info.get("ledger_totals")
    if not rows or totals is None:
        return False
    bold, body, small = fonts
    ws = wb.create_sheet("Ledger")
    _safe_cell(ws, 1, 1, "Every difference the published comparison counts",
               bold)
    _safe_cell(ws, 2, 1,
               f"{totals.data_rows:,} compared row(s) · "
               f"{totals.matched_rows:,} matched · "
               f"{totals.one_sided_rows:,} one-sided · "
               f"{totals.duplicate_groups:,} repeated-key group(s) covering "
               f"{totals.duplicate_member_rows:,} row(s)", small)
    _safe_cell(ws, 3, 1,
               f"Ledger digest {info.get('ledger_digest', '')} · reader v"
               f"{info.get('reader_version', '')} — this accounting is complete "
               "before any example is chosen; the images below illustrate it.",
               small)
    for col, label in enumerate(_LEDGER_HEADERS, start=1):
        _safe_cell(ws, 5, col, label, bold)
    r = 6
    for row in rows:
        for col, value in enumerate(row, start=1):
            _safe_cell(ws, r, col, value, body)
        _safe_cell(ws, r, len(_LEDGER_HEADERS), misses.get(row[0], ""), small)
        r += 1
    _safe_cell(ws, r + 1, 1, "Totals", bold)
    _safe_cell(ws, r + 1, 2, totals.difference_cells, bold)
    _safe_cell(ws, r + 1, 5, totals.context_cells, body)
    _safe_cell(ws, r + 1, 6, totals.equal_cells, body)
    _safe_cell(ws, r + 1, 7, totals.one_sided_cells, body)
    for col, width in (("A", 26), ("B", 20), ("C", 18), ("D", 26), ("E", 14),
                       ("F", 15), ("G", 15), ("H", 18), ("I", 70)):
        ws.column_dimensions[col].width = width
    return True


def _write_workbook(wb_path, img_dir, entries, misses, info,
                    layout=DEFAULT_LAYOUT, source_paths=(),
                    captured_sources=(), source_set_check=None,
                    commit_guard=None):
    """Write '<comparison> (evidence).xlsx' — a Summary sheet + one tab PER
    COMPARISON COLUMN for the chosen `layout` (side-by-side / stacked / both;
    'both' gives each column two tabs) — via a temp file + os.replace. Returns
    a short note when the previous workbook was locked open and the new one
    diverted to an exclusively reserved, unpredictable .new-<token> sibling."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    keys = _layout_keys(layout)                # the rendered layout(s), in order
    title = Font(name="Arial", size=14, bold=True)
    bold = Font(name="Arial", size=10, bold=True)
    body = Font(name="Arial", size=10)
    small = Font(name="Arial", size=9, color="666666")
    _safe_cell(ws, 1, 1, f"{info['report']} — visual evidence", title)
    _safe_cell(ws, 2, 1,
               f"Comparison: {info['comparison']}   ·   examples per column: "
               f"{info['examples']}   ·   sample seed: {info['seed']}", small)
    _safe_cell(ws, 3, 1,
               f"TSMIS PDFs: {info['tsmis_dir']}   ·   TSN PDFs: {info['tsn_dir']}",
               small)
    for col, value in enumerate(
            ("Column", "Route @ Post Mile", "TSMIS", "TSN", "Images",
             "Published cell", "Source rows"), start=1):
        _safe_cell(ws, 5, col, value, bold)
    r = 6
    for e in entries:
        _safe_cell(ws, r, 1, e["field"], body)
        _safe_cell(ws, r, 2, f"{e['route']} @ {e['key']}", body)
        _safe_cell(ws, r, 3, e["va"], body)
        _safe_cell(ws, r, 4, e["vb"], body)
        _safe_cell(ws, r, 5,
                   "  /  ".join(e[k] for k in keys if e.get(k)), body)
        # CMP-AUD-208: the published coordinates this image illustrates.
        _safe_cell(ws, r, 6, _published_cell_text(e), small)
        _safe_cell(ws, r, 7, _source_rows_text(e), small)
        r += 1
    for f, why in misses.items():
        _safe_cell(ws, r, 1, f, body)
        _safe_cell(ws, r, 2, f"no verifiable example — {why}", small)
        r += 1
    for col, width in (("A", 14), ("B", 24), ("C", 26), ("D", 26), ("E", 46),
                       ("F", 34), ("G", 26)):
        ws.column_dimensions[col].width = width

    used_sheet_names = {"Summary"}
    if _write_ledger_sheet(wb, info, misses, (bold, body, small)):
        used_sheet_names.add("Ledger")
    for img_key in keys:
        embed_w, label, layout_suffix = _LAYOUT_SPEC[img_key]
        _column_image_sheets(wb, entries, img_dir, img_key, embed_w, label,
                             (bold, small), used_sheet_names,
                             suffix=layout_suffix if len(keys) > 1 else "")

    tmp = None
    tmp_identity = ()

    def discard_tmp():
        if tmp is None or not os.path.lexists(tmp):
            return
        try:
            _ensure_captured_current(tmp_identity)
            artifact_store.ensure_outputs_do_not_alias_sources(
                (tmp,), source_paths, captured_sources=captured_sources)
            _require_output_guard(commit_guard, tmp,
                                  "temporary evidence-workbook cleanup")
        except (ValueError, owned_dir.OwnershipError):
            log.error("evidence: retained replaced/unsafe workbook temp instead "
                      "of deleting foreign or source content: %s", tmp)
            return
        Path(tmp).unlink(missing_ok=True)

    def discard_reserved_alt(alt, identity):
        if alt is None or not os.path.lexists(alt):
            return
        try:
            _ensure_captured_current(identity)
            _require_output_guard(commit_guard, alt,
                                  "fallback workbook-reservation cleanup")
        except (ValueError, owned_dir.OwnershipError):
            log.error("evidence: retained replaced fallback reservation instead "
                      "of deleting foreign content: %s", alt)
            return
        Path(alt).unlink(missing_ok=True)

    if source_set_check is not None:
        source_set_check()
    _require_output_guard(commit_guard, wb_path.parent,
                          "evidence workbook-folder creation")
    wb_path.parent.mkdir(parents=True, exist_ok=True)
    _require_output_guard(commit_guard, wb_path.parent,
                          "evidence workbook-folder creation")
    try:
        with tempfile.NamedTemporaryFile(
                mode="w+b", prefix=f".{wb_path.stem}.tmp-",
                suffix=wb_path.suffix, dir=wb_path.parent,
                delete=False) as temp_handle:
            tmp = Path(temp_handle.name)
            tmp_identity = artifact_store.capture_source_identities((tmp,))
            artifact_store.ensure_outputs_do_not_alias_sources(
                (wb_path, tmp), source_paths,
                captured_sources=captured_sources,
                require_sources_current=True)
            _require_output_guard(commit_guard, tmp,
                                  "temporary evidence-workbook write")
            wb.save(temp_handle)
    except BaseException:
        discard_tmp()
        raise
    try:
        if source_set_check is not None:
            source_set_check()
        artifact_store.ensure_outputs_do_not_alias_sources(
            (wb_path, tmp), source_paths,
            captured_sources=captured_sources, require_sources_current=True)
        _ensure_captured_current(tmp_identity)
        _require_output_guard(commit_guard, tmp,
                              "temporary evidence-workbook publish")
        _require_output_guard(commit_guard, wb_path,
                              "evidence-workbook publish")
        os.replace(tmp, wb_path)
        return None
    except ValueError:
        discard_tmp()
        raise
    except OSError as e:                       # workbook open in Excel, etc.
        log.warning("evidence: workbook swap failed: %s: %s",
                    type(e).__name__, e)
        alt = None
        alt_identity = ()
        try:
            # Reserve an unpredictable fallback exclusively. A foreign fixed
            # '<stem>.new.xlsx' sentinel is never named, replaced, or removed.
            with tempfile.NamedTemporaryFile(
                    mode="w+b", prefix=f"{wb_path.stem}.new-",
                    suffix=wb_path.suffix, dir=wb_path.parent,
                    delete=False) as alt_handle:
                alt = Path(alt_handle.name)
            alt_identity = artifact_store.capture_source_identities((alt,))
            if source_set_check is not None:
                source_set_check()
            artifact_store.ensure_outputs_do_not_alias_sources(
                (tmp, alt), source_paths, captured_sources=captured_sources,
                require_sources_current=True)
            _ensure_captured_current(tmp_identity)
            _ensure_captured_current(alt_identity)
            _require_output_guard(commit_guard, tmp,
                                  "temporary evidence-workbook fallback")
            _require_output_guard(commit_guard, alt,
                                  "fallback evidence-workbook publish")
            os.replace(tmp, alt)
            return f"previous evidence workbook is locked open — new set saved as {alt.name}"
        except ValueError:
            discard_reserved_alt(alt, alt_identity)
            discard_tmp()
            raise
        except OSError:
            discard_reserved_alt(alt, alt_identity)
            discard_tmp()
            return "previous evidence workbook is locked open — new workbook not saved"
        except BaseException:
            discard_reserved_alt(alt, alt_identity)
            discard_tmp()
            raise
    except BaseException:
        discard_tmp()
        raise


def _swap_dir(tmp_dir, target, source_paths=(), captured_sources=(),
              source_set_check=None, commit_guard=None,
              tmp_directory_identity=None):
    """Swap fresh images into place without touching guessed sibling names.

    The prior set is quarantined under an unpredictable per-operation name.
    Every failure after that rename restores it before propagating/returning;
    observed collisions are preserved and refused, never recursively removed.
    """
    tmp_dir, target = Path(tmp_dir), Path(target)
    old = _unique_dir_sibling(target, "old")
    tmp_identity = artifact_store.capture_source_identities((tmp_dir,))
    tmp_fs_identity = (tmp_directory_identity
                       if tmp_directory_identity is not None
                       else owned_dir.directory_identity(tmp_dir))
    if commit_guard is not None and tmp_fs_identity is None:
        raise owned_dir.OwnershipError(
            "The temporary evidence image folder changed before publication; "
            "it was left in place.")
    old_identity = ()
    old_fs_identity = None
    old_moved = False

    def ensure_swap_paths(*paths_to_check):
        if source_set_check is not None:
            source_set_check()
        artifact_store.ensure_outputs_do_not_alias_sources(
            paths_to_check, source_paths,
            directory_destinations=paths_to_check,
            captured_sources=captured_sources,
            require_sources_current=True)

    def restore_old():
        """Rollback only; a changed source-set must not block restoration."""
        nonlocal old_moved
        if not old_moved:
            return
        if not os.path.lexists(old):
            raise OSError(f"Evidence rollback lost its quarantine directory: {old}")
        if os.path.lexists(target):
            raise FileExistsError(
                f"Evidence rollback refused to replace a newly appeared path: {target}")
        _ensure_captured_current(old_identity)
        _require_output_guard(
            commit_guard, old, "evidence image-folder rollback",
            directory_identity=old_fs_identity)
        _require_output_guard(commit_guard, target,
                              "evidence image-folder rollback")
        # Do not require the compared sources to remain current for rollback:
        # that very guard may be why publication aborted. Still refuse a target
        # that now aliases the originally selected source object.
        artifact_store.ensure_outputs_do_not_alias_sources(
            (target,), source_paths, directory_destinations=(target,),
            captured_sources=captured_sources,
            require_sources_current=False)
        os.replace(old, target)
        old_moved = False
        _require_output_guard(
            commit_guard, target, "evidence image-folder rollback verification",
            directory_identity=old_fs_identity)

    def discard_old_after_success():
        nonlocal old_moved
        if not old_moved or not os.path.lexists(old):
            old_moved = False
            return
        try:
            _ensure_captured_current(old_identity)
            _require_output_guard(
                commit_guard, old, "prior evidence-folder cleanup",
                directory_identity=old_fs_identity)
            if not owned_dir.is_plain_directory_tree(old, old_fs_identity):
                raise ValueError(
                    "the prior evidence quarantine contains a linked or replaced path")
            shutil.rmtree(old)
            old_moved = False
        except (OSError, ValueError) as e:
            # Canonical publication succeeded. Retain an uncertain quarantine
            # instead of recursively deleting foreign replacement content.
            log.warning("evidence: retained prior image quarantine %s (%s: %s)",
                        old, type(e).__name__, e)

    try:
        ensure_swap_paths(tmp_dir, target, old)
        _ensure_captured_current(tmp_identity)
        _require_output_guard(
            commit_guard, tmp_dir, "temporary evidence-folder publication",
            directory_identity=tmp_fs_identity)
        _require_output_guard(commit_guard, target,
                              "evidence image-folder publication")
        _require_output_guard(commit_guard, old,
                              "evidence quarantine reservation")
        if os.path.lexists(old):
            raise FileExistsError(f"Evidence quarantine path appeared before use: {old}")
        if os.path.lexists(target):
            os.replace(target, old)
            old_moved = True
            old_identity = artifact_store.capture_source_identities((old,))
            old_fs_identity = owned_dir.directory_identity(old)
            if old_fs_identity is None:
                raise owned_dir.OwnershipError(
                    "The prior evidence image folder could not be bound after "
                    "quarantine; it was retained.")
            # Critical rollback boundary: every later guard failure is caught.
            ensure_swap_paths(tmp_dir, target, old)
            _ensure_captured_current(old_identity)
            _require_output_guard(
                commit_guard, old, "evidence quarantine verification",
                directory_identity=old_fs_identity)
            _require_output_guard(commit_guard, target,
                                  "evidence image-folder publication")
        _ensure_captured_current(tmp_identity)
        _require_output_guard(
            commit_guard, tmp_dir, "temporary evidence-folder publication",
            directory_identity=tmp_fs_identity)
        _require_output_guard(commit_guard, target,
                              "evidence image-folder publication")
        os.replace(tmp_dir, target)
        _require_output_guard(
            commit_guard, target, "evidence image-folder publish verification",
            directory_identity=tmp_fs_identity)
        discard_old_after_success()
        return None
    except OSError as e:                       # a file inside is locked open
        log.warning("evidence: image-folder swap failed: %s: %s",
                    type(e).__name__, e)
        try:
            alt = _unique_dir_sibling(target, "new")
            ensure_swap_paths(tmp_dir, alt)
            _ensure_captured_current(tmp_identity)
            _require_output_guard(
                commit_guard, tmp_dir, "temporary evidence-folder fallback",
                directory_identity=tmp_fs_identity)
            _require_output_guard(commit_guard, alt,
                                  "fallback evidence-folder publication")
            if os.path.lexists(alt):
                raise FileExistsError(f"Evidence fallback path appeared before use: {alt}")
            os.replace(tmp_dir, alt)
            _require_output_guard(
                commit_guard, alt, "fallback evidence-folder verification",
                directory_identity=tmp_fs_identity)
            restore_old()
            return f"previous images are locked open — new set saved to {alt.name}"
        except OSError:
            restore_old()
            return "previous images are locked open — new image set not saved"
        except BaseException:
            restore_old()
            raise
    except BaseException:
        # ValueError is the normal source-set/identity guard failure, but do the
        # same rollback for any unexpected guard exception.
        restore_old()
        raise


def _divert_images(tmp_dir, target, source_paths, captured_sources,
                   source_set_check, commit_guard, tmp_dir_fs_identity):
    """Move the freshly rendered images to an unpredictable .new-<token> sibling
    of `target` WITHOUT touching the canonical folder — used when the workbook
    could not reach its canonical name, so the OLD images stay in place and the
    workbook/images can't disagree (CMP-AUD-109). Returns the alt folder name,
    or None when even that could not be reserved (images dropped, keep-last-good;
    the temp is then cleaned by the caller)."""
    try:
        alt = _unique_dir_sibling(target, "new")
        if source_set_check is not None:
            source_set_check()
        artifact_store.ensure_outputs_do_not_alias_sources(
            (tmp_dir, alt), source_paths,
            directory_destinations=(tmp_dir, alt),
            captured_sources=captured_sources, require_sources_current=True)
        _require_output_guard(commit_guard, tmp_dir,
                              "temporary evidence-folder divert",
                              directory_identity=tmp_dir_fs_identity)
        _require_output_guard(commit_guard, alt,
                              "diverted evidence-folder publication")
        if os.path.lexists(alt):
            raise FileExistsError(f"Evidence divert path appeared before use: {alt}")
        os.replace(tmp_dir, alt)
        return alt.name
    except (OSError, ValueError, owned_dir.OwnershipError) as e:
        log.warning("evidence: could not divert images alongside the diverted "
                    "workbook: %s: %s", type(e).__name__, e)
        return None


def _quarantine_member(target, label, is_dir, source_paths, captured_sources,
                       source_set_check, commit_guard):
    """Move an existing canonical member out of the way under an unpredictable
    name, leaving its slot EMPTY so whatever replaces it stays undoable.

    Returns the quarantine path, or None when there was nothing there. Raises
    OSError when the member is locked open — the caller then publishes nothing,
    which is the same signal a direct replace would have given.
    """
    if not os.path.lexists(target):
        return None
    if source_set_check is not None:
        source_set_check()
    artifact_store.ensure_outputs_do_not_alias_sources(
        (target,), source_paths,
        directory_destinations=(target,) if is_dir else (),
        captured_sources=captured_sources, require_sources_current=True)
    quarantine = _unique_dir_sibling(target, "old", keep_suffix=not is_dir)
    _require_output_guard(commit_guard, target, f"evidence {label} quarantine")
    _require_output_guard(commit_guard, quarantine,
                          f"evidence {label} quarantine reservation")
    if os.path.lexists(quarantine):
        raise FileExistsError(
            f"Evidence quarantine path appeared before use: {quarantine}")
    os.replace(target, quarantine)
    return quarantine


def _restore_member(quarantine, target, label, commit_guard):
    """Put a quarantined member back. Rollback only, and loud when it fails —
    a lost quarantine is the one state that leaves nothing at the canonical
    name, so it may never be swallowed."""
    if quarantine is None:
        return
    if not os.path.lexists(quarantine):
        raise OSError(f"Evidence rollback lost its {label} quarantine: {quarantine}")
    if os.path.lexists(target):
        raise FileExistsError(
            f"Evidence rollback refused to replace a newly appeared path: {target}")
    _require_output_guard(commit_guard, target, f"evidence {label} rollback")
    os.replace(quarantine, target)


def _discard_quarantine(quarantine, label, is_dir, commit_guard):
    """Delete a superseded quarantine — never a path that replaced it."""
    if quarantine is None or not os.path.lexists(quarantine):
        return
    try:
        _require_output_guard(commit_guard, quarantine,
                              f"prior evidence {label} cleanup")
        if is_dir:
            fs_identity = owned_dir.directory_identity(quarantine)
            if not owned_dir.is_plain_directory_tree(quarantine, fs_identity):
                raise ValueError(
                    "the prior evidence quarantine contains a linked or replaced path")
            shutil.rmtree(quarantine)
        else:
            Path(quarantine).unlink(missing_ok=True)
    except (OSError, ValueError, owned_dir.OwnershipError) as e:
        log.warning("evidence: retained prior %s quarantine %s (%s: %s)",
                    label, quarantine, type(e).__name__, e)


def _withdraw_workbook(wb_path, commit_guard):
    """Take back a workbook committed into an emptied slot so the SET can roll
    back as a whole. It moves to a .new sibling rather than being deleted — the
    user still gets the freshly built workbook, just not at the canonical name."""
    alt = _unique_dir_sibling(wb_path, "new", keep_suffix=True)
    _require_output_guard(commit_guard, wb_path, "evidence workbook withdrawal")
    _require_output_guard(commit_guard, alt, "withdrawn evidence-workbook reservation")
    if os.path.lexists(alt):
        raise FileExistsError(f"Evidence withdrawal path appeared before use: {alt}")
    os.replace(wb_path, alt)
    return alt.name


def _commit_manifest(man_path, manifest, source_paths, captured_sources,
                     source_set_check, commit_guard):
    """Write the generation record into its EMPTIED slot — the last rename of
    the publication, so the manifest's presence means the whole set landed."""
    if source_set_check is not None:
        source_set_check()
    artifact_store.ensure_outputs_do_not_alias_sources(
        (man_path,), source_paths, captured_sources=captured_sources,
        require_sources_current=True)
    _require_output_guard(commit_guard, man_path.parent,
                          "evidence manifest-folder creation")
    man_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = None
    try:
        with tempfile.NamedTemporaryFile(
                mode="w", encoding="utf-8", newline="",
                prefix=f".{man_path.stem}.tmp-", suffix=man_path.suffix,
                dir=man_path.parent, delete=False) as handle:
            tmp = Path(handle.name)
            _require_output_guard(commit_guard, tmp,
                                  "temporary evidence-manifest write")
            handle.write(evidence_manifest.dumps(manifest))
        _require_output_guard(commit_guard, man_path,
                              "evidence manifest publication")
        os.replace(tmp, man_path)
        return None
    except (OSError, ValueError, owned_dir.OwnershipError) as e:
        if tmp is not None:
            Path(tmp).unlink(missing_ok=True)
        log.error("evidence: could not write the generation manifest %s: %s: %s",
                  man_path, type(e).__name__, e)
        return (f"the evidence generation record {man_path.name} could not be "
                "written — the published set is not durably identified")


def _publish_evidence_set(wb_path, img_dir, tmp_dir, entries, misses, info,
                          layout, source_paths, captured_sources,
                          source_set_check, commit_guard, tmp_dir_fs_identity,
                          man_path=None, manifest_for=None):
    """Publish the workbook + image folder + manifest as ONE set (CMP-AUD-109).

    Two-phase: every canonical member is QUARANTINED first, so each slot is
    empty and each replacement is undoable. Only then is the workbook written and
    the image folder swapped. If either fails — the common case is one of them
    locked open in Excel — the whole set rolls back to the prior generation and
    the new members divert to .new siblings. The canonical folder therefore never
    shows a new workbook beside old images, which the previous divert-both path
    could not guarantee because a committed workbook could not be taken back.

    The manifest commits last, into the slot emptied at the start, so its
    presence means the whole set landed and its absence never certifies a
    half-published one. A rollback restores the PRIOR manifest, which still
    describes the prior workbook and images exactly — the diverted state stays
    internally consistent.

    Returns {status, workbook, folder, manifest, note}: 'promoted' (all at
    canonical) or 'diverted' (see note). `workbook`/`folder` are the ACTUAL
    committed canonical paths, or None when nothing reached canonical.
    """
    man_quarantine = wb_quarantine = None
    images_handled = False
    try:
        if man_path is not None:
            man_quarantine = _quarantine_member(
                man_path, "manifest", False, source_paths, captured_sources,
                source_set_check, commit_guard)
        wb_quarantine = _quarantine_member(
            wb_path, "workbook", False, source_paths, captured_sources,
            source_set_check, commit_guard)
        wb_note = _write_workbook(
            wb_path, tmp_dir, entries, misses, info, layout=layout,
            source_paths=source_paths, captured_sources=captured_sources,
            source_set_check=source_set_check, commit_guard=commit_guard)
        if wb_note is None:
            dir_note = _swap_dir(
                tmp_dir, img_dir, source_paths=source_paths,
                captured_sources=captured_sources,
                source_set_check=source_set_check, commit_guard=commit_guard,
                tmp_directory_identity=tmp_dir_fs_identity)
            if dir_note is None:
                _discard_quarantine(wb_quarantine, "workbook", False, commit_guard)
                _discard_quarantine(man_quarantine, "manifest", False, commit_guard)
                note = None
                if man_path is not None and manifest_for is not None:
                    note = _commit_manifest(
                        man_path, manifest_for(wb_path, img_dir), source_paths,
                        captured_sources, source_set_check, commit_guard)
                return {"status": "promoted", "workbook": str(wb_path),
                        "folder": str(img_dir),
                        "manifest": str(man_path) if man_path else None,
                        "note": note}
            # The images are locked: take the workbook back so the canonical
            # pair stays the PRIOR consistent generation. `_swap_dir` has
            # already parked or dropped the new images, so they are not
            # diverted a second time.
            diverted_wb = _withdraw_workbook(wb_path, commit_guard)
            wb_note = (f"{dir_note}; the new workbook was withdrawn to "
                       f"{diverted_wb} so it cannot sit beside the old images")
            images_handled = True
    except BaseException:
        _restore_member(wb_quarantine, wb_path, "workbook", commit_guard)
        _restore_member(man_quarantine, man_path, "manifest", commit_guard)
        raise
    # Roll the whole set back to the prior generation and park the new one. The
    # restored manifest still describes the restored workbook and images
    # exactly, so the canonical set stays internally consistent.
    _restore_member(wb_quarantine, wb_path, "workbook", commit_guard)
    _restore_member(man_quarantine, man_path, "manifest", commit_guard)
    note = wb_note
    if not images_handled:
        dir_alt = _divert_images(tmp_dir, img_dir, source_paths,
                                 captured_sources, source_set_check,
                                 commit_guard, tmp_dir_fs_identity)
        note += (f"; new images saved to {dir_alt}" if dir_alt
                 else "; new images not saved")
    return {"status": "diverted", "workbook": None, "folder": None,
            "manifest": None, "note": note}
