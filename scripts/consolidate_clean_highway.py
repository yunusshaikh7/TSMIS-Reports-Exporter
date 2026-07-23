"""Build OUR "CA HIGHWAYS" clean-road workbook from the ArcGIS layer library.

The TSN `CA HIGHWAYS` extract (60,083 rows x 74 `THY_*` columns) is the vendor's
projection of the TSMIS ArcGIS LRS layers. This consolidator rebuilds the same
table INDEPENDENTLY from the owner's per-layer exports in `arcgis_layers/`, so
the ArcGIS tab can diff the two cell-for-cell. The model is the one value-proven
on route 001 (docs/planning/cleanroad-highways.md): select each layer's slices
as-of a chosen date (`LRSFromDate <= D < LRSToDate`; empty LRSToDate = current),
key everything on county + PM prefix + postmile (NEVER the odometers — three
calibrations exist), union every layer's breakpoints, and emit one row per
stretch with each layer's value painted across it.

Row structure follows the LRS alignments (measured 2026-07-22): as-of spans
always carry an explicit Alignment; LEFT-road attribute layers live wholly on
the L alignment (PM-coincident with R except on independent stretches); the
whole-road layers gain L spans only where alignments split. `SHS Highway Group`
is the structural driver — HG D/U segments are base rows; HG R / L segments are
the PM-suffixed independent-roadbed rows (the opposite side's block is null,
as TSN nulls it). HG coverage GAPS yield NO row: measured on the ORA 001
unconstructed gap (14.057–17.461), TSN skips the PM range too, and its 340
statewide HG=X inventory rows have no layer counterpart — they surface
one-sided in the comparison by design.

The output keeps the FULL 74-column THY header (owner decision 2026-07-22) in
three tiers: sourced (painted from a mapped layer), no-TSMIS-source (present +
empty + noted), and TSN-internal bookkeeping (present + empty + noted;
`THY_EXTRACT_DATE` alone is stamped with the build's as-of date). Every column
is indexed back to its source layer on the workbook's `Provenance` sheet — the
layer name + column plus the FeatureServer `Data Source` recorded by the
library's `00_INDEX.xlsx`.

Console-free; returns a ConsolidateResult. openpyxl loads lazily.
"""
from __future__ import annotations

import logging
import os
import re
from collections import defaultdict
from pathlib import Path

import city_codes
import clean_highway_columns as chc
import clean_road_layers as crl
import consolidation_meta
import outcome
import paths
from events import ConsolidateResult, Events

log = logging.getLogger(__name__)

REPORT_NAME = "Clean Road Highway (ArcGIS)"
SHEET_NAME = chc.ARC_SHEET
MARKER_SHEET = chc.ARC_MARKER_SHEET
BUILD_VERSION = 1
FILENAME = "clean_highway_built.xlsx"
OUT_DIR = paths.OUTPUT_ROOT / "arcgis_cleanroad"
OUT_PATH = OUT_DIR / FILENAME

# The 74-column THY header (clean_highway_columns is the shared contract).
THY_HEADER = list(chc.HEADER)
_COL = {name: i for i, name in enumerate(THY_HEADER)}

# The identity/span columns every span layer contributes. District rides every
# span (labels dialect), so THY_DISTRICT_CODE paints from the covering spans
# themselves — the SHS District layer's own spans are route-length and lose
# their middle counties to the two-ended split. RouteID drives the SHS filter
# on the all-roads layers (City).
_SPAN_ID = ["District", "RouteNum", "RouteSuffix", "Alignment", "BeginCounty",
            "EndCounty", "BeginPMPrefix", "BeginPMMeasure", "EndPMPrefix",
            "EndPMMeasure", "BeginODMeasure", "EndODMeasure",
            "InventoryItemStartDate", "LRSFromDate", "LRSToDate", "RouteID"]
_POINT_ID = ["RouteNum", "RouteSuffix", "Alignment", "County", "PMPrefix",
             "PMMeasure", "LRSFromDate", "LRSToDate"]

# tag -> (layer name, attribute columns, family). Families: "whole" paints base
# rows AND its own alignment's independent rows; "right"/"left" paint that
# roadbed's block; "tvs" is the ADT interpolation source. (City is deliberately
# NOT read: its City_Code carries city NAMES where TSN carries TASAS city
# letter codes — no code table exists in the library, so THY_CITY_CODE is a
# noted no-source column rather than 23k fabricated differences. County Code
# is likewise unread for highways: county identity/extents come from every
# span's own Begin/End county fields.)
SPAN_LAYERS = {
    "HG": ("SHS Highway Group", ["Highway_Group"], "whole"),
    "MED": ("SHS Median", ["Median_Type", "Median_Width", "Median_Variance"],
            "whole"),
    "BAR": ("SHS Barrier", ["Barrier_Type"], "whole"),
    "CURB": ("SHS Curb Landscape", ["Curb_Landscape"], "whole"),
    "ACC": ("SHS Access Control", ["SHS_Access_Control"], "whole"),
    "TER": ("Terrain Type", ["Terrain_Type"], "whole"),
    "DSP": ("SHS Design Speed", ["Design_Speed"], "whole"),
    "POP": ("SHS Population", ["Population_Code"], "whole"),
    "NON": ("SHS Non Add Mileage", ["Non_Add_Mileage"], "whole"),
    # City cuts rows at every SHS city limit AND paints THY_CITY_CODE — the
    # layer carries city NAMES; city_codes.norm_city translates them to the
    # TASAS letter codes (the table was derived from statewide co-location
    # with the extract; an unmapped name passes through verbatim so it
    # surfaces in the comparison instead of vanishing).
    "CITY": ("City", ["City_Code"], "whole"),
    "TOLL": ("SHS Tolls", ["Toll_Type"], "whole"),
    "FOR": ("SHS Forest HWY", ["Forest_Hwy"], "whole"),
    "NET": ("SHS Inv Network Date", ["Network_Start_Date", "SegOrderId"],
            "whole"),
    "TVS": ("Traffic Volume Segments",
            ["AADT", "AADT_YEAR", "AADT_AHEAD", "AADT_BACK"], "tvs"),
    "TWR": ("SHS Travel Way R", ["Travel_Way_Width_R", "Total_Num_Lanes_R"],
            "right"),
    "SUR": ("SHS Surface Type R", ["Surface_Type_R"], "right"),
    "SPR": ("SHS Special Feature R", ["Special_Feature_Type_R"], "right"),
    "OSR": ("SHS O Shld Width R",
            ["Shld_Width_Total_Out_R", "Shld_Width_Treated_Out_R"], "right"),
    "ISR": ("SHS I Shld Width R",
            ["Shld_Width_Total_In_R", "Shld_Width_Treated_In_R"], "right"),
    "TWL": ("SHS Travel Way L", ["Travel_Way_Width_L", "Total_Num_Lanes_L"],
            "left"),
    "SUL": ("SHS Surface Type L", ["Surface_Type_L"], "left"),
    "SPL": ("SHS Special Feature L", ["Special_Feature_Type_L"], "left"),
    "OSL": ("SHS O Shld Width L",
            ["Shld_Width_Total_Out_L", "Shld_Width_Treated_Out_L"], "left"),
    "ISL": ("SHS I Shld Width L",
            ["Shld_Width_Total_In_L", "Shld_Width_Treated_In_L"], "left"),
}
POINT_LAYERS = {
    "LMK": ("SHS Landmark", ["Landmarks_Short"]),
    "EQP": ("Equation Points", ["hslDescription"]),
    "RBR": ("SHS Route Break", ["Route_Break_Type"]),
}
_LEFT_TAGS = tuple(t for t, s in SPAN_LAYERS.items() if s[2] == "left")
_RIGHT_TAGS = tuple(t for t, s in SPAN_LAYERS.items() if s[2] == "right")

# Every layer this consolidator streams — all must be staged or the build
# refuses (naming exactly what's missing).
HIGHWAY_LAYERS = tuple(sorted(
    {spec[0] for spec in SPAN_LAYERS.values()}
    | {spec[0] for spec in POINT_LAYERS.values()}))

# Roadbed blocks: the opposite independent row's block stays null (TSN nulls
# the LT block on R rows and the RT block on L rows).

_ROUTE_TOKEN_RE = re.compile(r"^(\d{3})([SU]?)$")


class _Span:
    __slots__ = ("b", "e", "vals", "rank", "item", "od", "od_len", "district")

    def __init__(self, b, e, vals, rank, item, od, od_len=None, district=""):
        self.b, self.e, self.vals = b, e, vals
        self.rank, self.item, self.od = rank, item, od
        self.od_len = od_len
        self.district = district


def _route_token(route_num, route_suffix):
    return crl.norm_route(route_num) + crl.dot_none(route_suffix).strip().upper()


def _cancelled():
    return ConsolidateResult(
        status="cancelled", message="Cancelled. Nothing was written.",
        completion=outcome.CANCELLED)


# --------------------------------------------------------------------------- #
# reading
# --------------------------------------------------------------------------- #
def _read_span_layer(lib, index, tag, *, asof, events, buckets, parked):
    """Stream one span layer's as-of slices into `buckets[(route, align,
    county, prefix)][tag]`; cross-county/prefix spans park for the split pass."""
    layer, attrs, family = SPAN_LAYERS[tag]
    entry = index.get(layer) or {}
    n_asof = n_cross = 0
    for r in crl.stream_layer(lib["present"][layer], _SPAN_ID + attrs,
                              layer_name=layer, expected_rows=entry.get("rows"),
                              optional=("InventoryItemStartDate",)):
        rid = r["RouteID"]
        if isinstance(rid, str) and rid and not rid.startswith("SHS_"):
            continue                     # all-roads layers: SHS routes only
        if not crl.is_asof(crl.to_serial(r["LRSFromDate"]),
                           crl.to_serial(r["LRSToDate"]), asof):
            continue
        b, e = crl.pm_units(r["BeginPMMeasure"]), crl.pm_units(r["EndPMMeasure"])
        if b is None or e is None:
            continue
        n_asof += 1
        route = _route_token(r["RouteNum"], r["RouteSuffix"])
        align = crl.norm_alignment(r["Alignment"]) or "R"
        c1, c2 = crl.norm_county(r["BeginCounty"]), crl.norm_county(r["EndCounty"])
        p1 = crl.dot_none(r["BeginPMPrefix"]).strip().upper()
        p2 = crl.dot_none(r["EndPMPrefix"]).strip().upper()
        item = crl.to_serial(r["InventoryItemStartDate"])
        if family == "tvs":
            rank = (crl.to_serial(r["AADT_YEAR"]) or 0.0,
                    crl.to_serial(r["LRSFromDate"]) or 0.0)
        else:
            rank = (crl.to_serial(r["LRSFromDate"]) or 0.0, item or 0.0)
        od_b = crl.to_serial(r["BeginODMeasure"])
        od_e = crl.to_serial(r["EndODMeasure"])
        od_len = (od_e - od_b) if od_b is not None and od_e is not None else None
        span = _Span(b, e, tuple(r[a] for a in attrs), rank, item, od_b, od_len,
                     crl.norm_district(r["District"]))
        if c1 == c2 and p1 == p2:
            if e <= b:
                continue                     # degenerate within one PM space
            buckets[(route, align, c1, p1)][tag].append(span)
        else:
            # A cross-county/prefix span's end PM lives in the NEXT space, so
            # e < b is the NORMAL shape here — park it for the chain split.
            n_cross += 1
            parked.append((tag, route, align, c1, p1, c2, p2, span))
    events.on_log(f"  {layer}: {n_asof:,} as-of spans"
                  + (f" ({n_cross} cross-county)" if n_cross else ""))


def _read_point_layer(lib, index, tag, *, asof, events, points):
    """Stream one point layer's as-of rows into `points[(route, county,
    prefix)][tag]` = {pm_units: [values]} (several points can share a PM)."""
    layer, attrs = POINT_LAYERS[tag]
    entry = index.get(layer) or {}
    n_asof = 0
    for r in crl.stream_layer(lib["present"][layer], _POINT_ID + attrs,
                              layer_name=layer, expected_rows=entry.get("rows")):
        # Equation points are PERMANENT PM-line facts: the layer stamps its
        # own LOAD date as LRSFromDate (measured: all 1,570 say 2026-01-14),
        # so an as-of filter on it would erase every equate. Only a real
        # retirement (LRSToDate) ends one; the other point layers keep the
        # full as-of rule.
        f = (None if tag == "EQP" else crl.to_serial(r["LRSFromDate"]))
        if not crl.is_asof(f if f is not None else -1e9,
                           crl.to_serial(r["LRSToDate"]), asof):
            continue
        pm = crl.pm_units(r["PMMeasure"])
        if pm is None:
            continue
        n_asof += 1
        route = _route_token(r["RouteNum"], r["RouteSuffix"])
        county = crl.norm_county(r["County"])
        prefix = crl.dot_none(r["PMPrefix"]).strip().upper()
        points[(route, county, prefix)].setdefault(tag, {}).setdefault(
            pm, []).append(r[attrs[0]])
    events.on_log(f"  {layer}: {n_asof:,} as-of points")


def _split_parked(parked, buckets, warnings):
    """Split cross-county/prefix spans along the county CHAIN they traverse.

    Long administrative spans (Access Control, Terrain, the district-scale
    layers) cross SEVERAL county lines; a two-ended split loses every middle
    county, which surfaced as attribute holes across whole counties. The walk:
    the first county keeps [span begin → its own extent end], every chain
    county whose HG odometer window falls inside the span's odometer range is
    covered WHOLLY (its full extent), and the last county takes [its origin →
    span end] (a continuation county's postmiles start at 0 — measured: MON
    101.178 → SCR 0.043 covers SCR from 0.000). Odometers order and APPORTION
    here, never join. A span with no other resolution drops with a warning —
    never guessed."""
    extents = {}
    od_windows = {}
    for (route, _align, county, prefix), tags in buckets.items():
        key = (route, county, prefix)
        lo, hi = extents.get(key, (None, None))
        for spans in tags.values():
            for s in spans:
                lo = s.b if lo is None else min(lo, s.b)
                hi = s.e if hi is None else max(hi, s.e)
        extents[key] = (lo, hi)
        hg = tags.get("HG")
        if hg:
            wins = od_windows.setdefault(key, [])
            wins.extend((s.od, s.od + s.od_len) for s in hg
                        if s.od is not None and s.od_len is not None)
    od_windows = {k: _merge_windows(v) for k, v in od_windows.items()}

    for tag, route, align, c1, p1, c2, p2, span in parked:
        placed = False
        if span.e == 0 and span.od_len is not None:
            # The span ENDS exactly at the next space's origin — the second
            # part is zero-length by definition, so the whole odometer length
            # belongs to the first county.
            hi1 = span.b + int(round(span.od_len * crl.PM_SCALE))
            if span.b < hi1:
                buckets[(route, align, c1, p1)][tag].append(
                    _Span(span.b, hi1, span.vals, span.rank, span.item,
                          span.od, span.od_len, span.district))
                continue
        hi1 = (extents.get((route, c1, p1)) or (None, None))[1]
        if hi1 is None and span.od_len is not None:
            hi1 = span.b + max(int(round(span.od_len * crl.PM_SCALE))
                               - span.e, 0)
        if hi1 is not None and span.b < hi1:
            buckets[(route, align, c1, p1)][tag].append(
                _Span(span.b, hi1, span.vals, span.rank, span.item,
                      span.od, (hi1 - span.b) / crl.PM_SCALE, span.district))
            placed = True
        if 0 < span.e:
            buckets[(route, align, c2, p2)][tag].append(
                _Span(0, span.e, span.vals, span.rank, span.item,
                      (span.od + span.od_len - span.e / crl.PM_SCALE
                       if span.od is not None and span.od_len is not None
                       else None),
                      span.e / crl.PM_SCALE, span.district))
            placed = True
        # Middle counties: chain cps whose odometer window sits inside the
        # span's odometer range get the span across their WHOLE extent.
        if span.od is not None and span.od_len is not None:
            od_lo, od_hi = span.od, span.od + span.od_len
            for (rt, county, prefix), wins in od_windows.items():
                if rt != route or (county, prefix) in ((c1, p1), (c2, p2)):
                    continue
                total = sum(we - wb for wb, we in wins)
                if total <= 0:
                    continue
                inside = _overlap_len((od_lo, od_hi), wins)
                if inside < 0.5 * total:
                    continue
                lo, hi = extents.get((rt, county, prefix), (None, None))
                if lo is None or hi is None or hi <= lo:
                    continue
                buckets[(route, align, county, prefix)][tag].append(
                    _Span(lo, hi, span.vals, span.rank, span.item,
                          wins[0][0], (hi - lo) / crl.PM_SCALE, span.district))
                placed = True
        if not placed:
            warnings.append(
                f"{SPAN_LAYERS[tag][0]}: a cross-county span "
                f"({route} {c1}/{p1 or '-'} {crl.pm_text(span.b)} → {c2}/"
                f"{p2 or '-'} {crl.pm_text(span.e)}) could not be split — "
                "skipped.")


# --------------------------------------------------------------------------- #
# geometry helpers
# --------------------------------------------------------------------------- #
def _clip_spans(spans, windows, *, inside):
    """Clip `[_Span]` to (inside=True) or away from (inside=False) the sorted,
    disjoint `windows` [(b, e)]."""
    if not windows:
        return list(spans) if not inside else []
    out = []
    for s in spans:
        keep = []
        if inside:
            for wb, we in windows:
                b, e = max(s.b, wb), min(s.e, we)
                if b < e:
                    keep.append((b, e))
        else:
            cur = s.b
            for wb, we in windows:
                if we <= cur:
                    continue
                if wb >= s.e:
                    break
                if cur < wb:
                    keep.append((cur, min(wb, s.e)))
                cur = max(cur, we)
            if cur < s.e:
                keep.append((cur, s.e))
        for b, e in keep:
            out.append(_Span(b, e, s.vals, s.rank, s.item, s.od)
                       if (b, e) != (s.b, s.e) else s)
    return out


def _merge_windows(intervals):
    """Sorted disjoint union of [(b, e)]."""
    out = []
    for b, e in sorted(intervals):
        if out and b <= out[-1][1]:
            out[-1] = (out[-1][0], max(out[-1][1], e))
        else:
            out.append((b, e))
    return out


# --------------------------------------------------------------------------- #
# painting
# --------------------------------------------------------------------------- #
def _seg_code(seg, tag, pos=0):
    span = seg.get(tag)
    if span is None:
        return None
    return crl.code_of(span.vals[pos])


def _seg_num(seg, tag, pos=0):
    span = seg.get(tag)
    if span is None:
        return None
    v = span.vals[pos]
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v).strip()
    return int(f) if f == int(f) else f


def _block_eff_date(seg, tags):
    """The composite block effective date: the OLDEST member layer's
    InventoryItemStartDate over the covering spans (measured on route 001 —
    newest lost to 1964-domain dates; single layers agree only 82–96%, the
    block date is a composite and this is the closer candidate)."""
    best = None
    for tag in tags:
        span = seg.get(tag)
        if span is not None and span.item is not None:
            best = span.item if best is None else min(best, span.item)
    return crl.serial_to_date(best)


def _seg_district(seg):
    """The segment's district from its covering spans' own District identity
    (the HG span first — every span carries District in this dialect)."""
    hg = seg.get("HG")
    if hg is not None and hg.district:
        return hg.district
    for span in seg.values():
        if isinstance(span, _Span) and span.district:
            return span.district
    return ""


def _pick_landmark(values):
    """One THY landmark text from the point(s) at a row's begin: the longest
    non-blank (co-located structural markers like 'BEGIN REALIGNMENT' ride
    beside the substantive landmark text)."""
    texts = [str(v) for v in values if v is not None and str(v).strip()]
    return max(texts, key=len) if texts else None


def _paint_row(route, county, prefix, b, e, seg, kind, pts, asof_date):
    """One output row (a 74-slot list) for segment [b, e) of `kind` in
    {'base', 'R', 'L'}. ADT/offsets/break-desc are finalized later."""
    m = _ROUTE_TOKEN_RE.match(route)
    row = [None] * len(THY_HEADER)

    def put(name, value):
        row[_COL[name]] = value

    put("THY_DISTRICT_CODE", _seg_district(seg))
    put("THY_COUNTY_CODE", county)
    put("THY_ROUTE_NAME", m.group(1) if m else route)
    put("THY_ROUTE_SUFFIX_CODE", (m.group(2) or None) if m else None)
    put("THY_PM_PREFIX_CODE", prefix or None)
    put("THY_BEGIN_PM_AMT", crl.pm_float(b))
    put("THY_END_PM_AMT", crl.pm_float(e))
    put("THY_PM_SUFFIX_CODE", kind if kind in ("R", "L") else None)
    put("THY_LENGTH_MILES_AMT", round((e - b) / crl.PM_SCALE, 5))
    put("THY_HIGHWAY_GROUP_CODE", _seg_code(seg, "HG"))
    put("THY_EXTRACT_DATE", asof_date)

    put("THY_MEDIAN_EFF_DATE", _block_eff_date(seg, ("MED", "BAR", "CURB")))
    put("THY_MEDIAN_TYPE_CODE", _seg_code(seg, "MED", 0))
    put("THY_MEDIAN_WIDTH_AMT", _seg_num(seg, "MED", 1))
    put("THY_MEDIAN_WIDTH_VAR_CODE", _seg_code(seg, "MED", 2))
    put("THY_MEDIAN_BARRIER_CODE", _seg_code(seg, "BAR"))
    put("THY_CURB_LANDSCAPE_CODE", _seg_code(seg, "CURB"))
    put("THY_HIGHWAY_ACCESS_CODE", _seg_code(seg, "ACC"))
    put("THY_ACCESS_EFF_DATE",
        crl.serial_to_date(seg["ACC"].item) if seg.get("ACC") else None)
    put("THY_TERRAIN_CODE", _seg_code(seg, "TER"))
    put("THY_DESIGN_SPEED_AMT", _seg_num(seg, "DSP"))
    put("THY_POPULATION_CODE", _seg_code(seg, "POP"))
    city = seg.get("CITY")
    if city is not None:
        put("THY_CITY_CODE", city_codes.norm_city(city.vals[0]) or None)
    non = _seg_code(seg, "NON")
    put("THY_NON_ADD_CODE", non if non in ("N", "A") else "A")
    put("THY_TOLL_FOREST_CODE",
        1 if seg.get("TOLL") is not None else 2 if seg.get("FOR") is not None
        else 0)
    net = seg.get("NET")
    if net is not None:
        put("THY_RECORD_DATE", crl.serial_to_date(crl.to_serial(net.vals[0])))
        put("THY_SEG_ORDER_ID", _seg_num(seg, "NET", 1))

    if kind in ("base", "R"):
        put("THY_RIGHT_ROAD_EFF_DATE", _block_eff_date(seg, _RIGHT_TAGS))
        put("THY_RT_SURF_TYPE_CODE", _seg_code(seg, "SUR"))
        put("THY_RT_TRAV_WAY_WIDTH_AMT", _seg_num(seg, "TWR", 0))
        put("THY_RT_LANES_AMT", _seg_num(seg, "TWR", 1))
        put("THY_RT_SPEC_FEATURES_CODE", _seg_code(seg, "SPR"))
        put("THY_RT_O_SHD_TOT_WIDTH_AMT", _seg_num(seg, "OSR", 0))
        put("THY_RT_O_SHD_TRT_WIDTH_AMT", _seg_num(seg, "OSR", 1))
        put("THY_RT_I_SHD_TOT_WIDTH_AMT", _seg_num(seg, "ISR", 0))
        put("THY_RT_I_SHD_TRT_WIDTH_AMT", _seg_num(seg, "ISR", 1))
    if kind in ("base", "L"):
        put("THY_LEFT_ROAD_EFF_DATE", _block_eff_date(seg, _LEFT_TAGS))
        put("THY_LT_SURF_TYPE_CODE", _seg_code(seg, "SUL"))
        put("THY_LT_TRAV_WAY_WIDTH_AMT", _seg_num(seg, "TWL", 0))
        put("THY_LT_LANES_AMT", _seg_num(seg, "TWL", 1))
        put("THY_LT_SPEC_FEATURES_CODE", _seg_code(seg, "SPL"))
        put("THY_LT_O_SHD_TOT_WIDTH_AMT", _seg_num(seg, "OSL", 0))
        put("THY_LT_O_SHD_TRT_WIDTH_AMT", _seg_num(seg, "OSL", 1))
        put("THY_LT_I_SHD_TOT_WIDTH_AMT", _seg_num(seg, "ISL", 0))
        put("THY_LT_I_SHD_TRT_WIDTH_AMT", _seg_num(seg, "ISL", 1))

    lmk = _pick_landmark(pts.get("LMK", {}).get(b, ()))
    if lmk is not None:
        put("THY_LANDMARK_SHORT_DESC", lmk)
    at_equate = b in pts.get("EQP", {})
    if at_equate:
        put("THY_EQUATE_CODE", "E")
    rbr = pts.get("RBR", {}).get(b)
    if rbr and not at_equate:
        # A break/resume point that coincides with an equation point reads as
        # the equate (THY leaves BREAK_DESC empty there).
        put("THY_BREAK_DESC",
            "U-BR" if "RESUME" in str(rbr[0]).upper() else "R-BR")

    tvs = seg.get("TVS")
    hg = seg.get("HG")
    if hg is not None and hg.od is not None:
        od = hg.od + (b - hg.b) / crl.PM_SCALE      # ordering only, never a value
    else:
        od = min((s.od for s in seg.values()
                  if isinstance(s, _Span) and s.od is not None), default=None)
    return {"row": row, "b": b, "e": e, "kind": kind, "county": county,
            "prefix": prefix, "tvs": tvs, "od": od}


# --------------------------------------------------------------------------- #
# the ADT interpolation family (Traffic Volume Segments)
# --------------------------------------------------------------------------- #
def _num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):  # silent-ok: a non-numeric count cell contributes no ADT arithmetic; the painted cell stays empty and the comparison shows it
        return None


def _apply_adt(rows):
    """Paint THY_PROFILE_CODE / THY_ADT_AMT / THY_CHANGE_PER_MILE_AMT along
    ordered rows: a row starting at its covering TVS span's begin on a
    contiguous stretch is a profile anchor (P); following rows are S rows
    stepped at the span's slope times the previous row's length, rounded like
    TSN's own arithmetic (proven on route 001: +41/+34/+27 per row = chg ×
    row length). A row entering a span mid-way interpolates from the span's
    begin instead."""
    prev = None                            # the previous OUTPUT row (any kind)
    for rec in rows:
        row, tvs = rec["row"], rec["tvs"]
        if tvs is None:
            row[_COL["THY_PROFILE_CODE"]] = "S"
            prev = rec
            continue
        length_mi = (tvs.e - tvs.b) / crl.PM_SCALE
        ahead = _num(tvs.vals[2]) or 0.0
        back = _num(tvs.vals[3]) or 0.0
        # Measured against THY on route 001 (three anchors triangulated it):
        # the value at a span's BEGIN is its AADT_AHEAD count, the value at
        # its END is AADT_BACK, so the along-PM slope is (BACK − AHEAD) /
        # span length. AADT itself is the station's midpoint count and sits a
        # constant offset off both — anchoring on it shifted every row.
        slope = (back - ahead) / length_mi if length_mi else 0.0
        anchor = _num(tvs.vals[2])           # AADT_AHEAD = the begin value
        anchored = rec["b"] == tvs.b
        # Interpolate every row from the SPAN'S OWN anchor (never chain off
        # the previous row's already-rounded value — chained rounding drifted
        # one count high against THY within a few rows).
        if anchor is not None:
            adt = anchor + slope * (rec["b"] - tvs.b) / crl.PM_SCALE
        else:
            adt = None
        # P marks a profile station the row STARTS at: a contiguous handoff
        # (the previous row touches this one) or the route's first coverage.
        # A span that begins across a PM gap is NOT a station row (THY says S
        # at the ORA 17.461 gap re-entry).
        contiguous = prev is None or prev["e"] == rec["b"]
        row[_COL["THY_PROFILE_CODE"]] = ("P" if anchored and contiguous
                                         else "S")
        row[_COL["THY_ADT_AMT"]] = int(round(adt)) if adt is not None else None
        row[_COL["THY_CHANGE_PER_MILE_AMT"]] = round(slope, 4)
        prev = rec


# --------------------------------------------------------------------------- #
# per-route assembly
# --------------------------------------------------------------------------- #
def _reduce_tvs(spans):
    """Reduce the Traffic Volume spans to the WINNING profile per stretch
    (latest AADT_YEAR, then newest slice) BEFORE the main overlay — three
    year-vintages are LRS-live concurrently, and feeding every vintage's edges
    into the breakpoint union fabricates row cuts THY doesn't have. Returns
    clipped (b, e, original_span, rank) tuples: the geometry is the winning
    region, the payload keeps the ORIGINAL span (its own begin anchors the
    profile arithmetic)."""
    if not spans:
        return []
    segs = crl.overlay({"T": [(s.b, s.e, s, s.rank) for s in spans]})
    out = []
    for b, e, seg in segs:
        span = seg.get("T")
        if span is None:
            continue
        if out and out[-1][2] is span and out[-1][1] == b:
            out[-1] = (out[-1][0], e, span, span.rank)
        else:
            out.append((b, e, span, span.rank))
    return out


def _overlap_len(window, intervals):
    b, e = window
    total = 0.0
    for wb, we in intervals:
        lo, hi = max(b, wb), min(e, we)
        if lo < hi:
            total += hi - lo
    return total


def _build_cp(route, county, prefix, rbucket, lbucket, pts, asof_date):
    """All rows for one (route, county, prefix): base rows outside the
    independent windows, R/L rows inside them. HG coverage GAPS yield NO row
    at all — measured on the ORA 001 unconstructed gap, TSN skips the PM
    range entirely (its 340 statewide HG=X inventory rows have no layer
    counterpart and surface one-sided in the comparison, by design)."""
    hg_l = lbucket.get("HG", [])
    win_l = _merge_windows([(s.b, s.e) for s in hg_l
                            if crl.code_of(s.vals[0]) == "L"])
    cuts = sorted({pm for tag in ("LMK", "EQP", "RBR")
                   for pm in pts.get(tag, {})})

    # The base/R universe: everything on the R alignment, plus the left-side
    # layers (their L-alignment PMs coincide with R off the independent
    # stretches) clipped away from the L windows.
    spans_r = {tag: [(s.b, s.e, s, s.rank) for s in spans]
               for tag, spans in rbucket.items() if tag != "TVS"}
    if rbucket.get("TVS"):
        spans_r["TVS"] = _reduce_tvs(rbucket["TVS"])
    for tag in _LEFT_TAGS:
        clipped = _clip_spans(lbucket.get(tag, []), win_l, inside=False)
        if clipped:
            spans_r[tag] = [(s.b, s.e, s, s.rank) for s in clipped]
    rows = []
    for b, e, seg in crl.overlay(spans_r, cuts=cuts):
        hg = seg.get("HG")
        if hg is None:
            continue
        kind = "R" if crl.code_of(hg.vals[0]) == "R" else "base"
        rows.append(_paint_row(route, county, prefix, b, e, seg, kind, pts,
                               asof_date))

    # The L universe: the L-alignment whole-road spans exist only on the
    # independent stretches; left-side layers clipped INTO those windows.
    spans_l = {tag: [(s.b, s.e, s, s.rank) for s in spans]
               for tag, spans in lbucket.items() if tag not in _LEFT_TAGS}
    for tag in _LEFT_TAGS:
        clipped = _clip_spans(lbucket.get(tag, []), win_l, inside=True)
        if clipped:
            spans_l[tag] = [(s.b, s.e, s, s.rank) for s in clipped]
    for b, e, seg in crl.overlay(spans_l, cuts=cuts):
        hg = seg.get("HG")
        if hg is None or crl.code_of(hg.vals[0]) != "L":
            continue
        rows.append(_paint_row(route, county, prefix, b, e, seg, "L", pts,
                               asof_date))
    return rows


_BREAK_PRIORITY = ("BEG", "END", "D-C", "CNTY", "DIST", "R-BR", "U-BR",
                   "BEGL", "ENDR")


def _finalize_route(recs):
    """Order one route's rows, then stamp the running offsets and the
    structural THY_BREAK_DESC markers.

    Offsets are the PM-CONTINUED CUMULATIVE (measured on route 001, three
    behaviors triangulated): a running sum of row lengths, seeded with the
    route's first begin PM, that additionally JUMPS FORWARD by any positive
    PM gap between consecutive rows (an unconstructed stretch's distance is
    carried; the offset at the ORA 17.461 gap re-entry is 17.461). A
    NEGATIVE PM step — a county line's reset to 0, or a prefix space
    re-entering slightly behind (plain 18.011 → R 18.000 continues at
    18.011) — never adjusts the cumulative."""
    def put(rec, name, value):
        rec["row"][_COL[name]] = value

    cum = None
    prev_end_pm = None
    for i, rec in enumerate(recs):
        row = rec["row"]
        begin_pm = crl.pm_float(rec["b"])
        if rec["kind"] == "L":
            # The left independent roadbed is PARALLEL mileage: it reads the
            # cumulative where it stands but never advances the corridor line.
            base = begin_pm if cum is None else cum
            row[_COL["THY_BEGIN_OFFSET_AMT"]] = round(base, 5)
            row[_COL["THY_END_OFFSET_AMT"]] = round(
                base + row[_COL["THY_LENGTH_MILES_AMT"]], 5)
        else:
            if cum is None:
                cum = begin_pm
            elif prev_end_pm is not None and begin_pm > prev_end_pm:
                cum = round(cum + (begin_pm - prev_end_pm), 5)
            row[_COL["THY_BEGIN_OFFSET_AMT"]] = round(cum, 5)
            cum = round(cum + row[_COL["THY_LENGTH_MILES_AMT"]], 5)
            row[_COL["THY_END_OFFSET_AMT"]] = cum
            prev_end_pm = crl.pm_float(rec["e"])

        marks = []
        if i == 0:
            marks.append("BEG")
        if i == len(recs) - 1:
            marks.append("END")
        prev = recs[i - 1] if i else None
        if prev is not None and rec["county"] != prev["county"]:
            same_dist = (row[_COL["THY_DISTRICT_CODE"]]
                         == prev["row"][_COL["THY_DISTRICT_CODE"]])
            marks.append("CNTY" if same_dist else "D-C")
        elif prev is not None and (row[_COL["THY_DISTRICT_CODE"]]
                                   != prev["row"][_COL["THY_DISTRICT_CODE"]]):
            marks.append("DIST")
        if rec["kind"] == "L" and (prev is None or prev["kind"] != "L"):
            marks.append("BEGL")
        nxt = recs[i + 1] if i + 1 < len(recs) else None
        if rec["kind"] == "R" and (nxt is None or nxt["kind"] != "R"):
            marks.append("ENDR")
        if marks:
            existing = row[_COL["THY_BREAK_DESC"]]
            if existing:
                marks.append(existing)
            best = min(marks, key=lambda m: (_BREAK_PRIORITY.index(m)
                                             if m in _BREAK_PRIORITY else 99))
            put(rec, "THY_BREAK_DESC", best)


def _build_route(route, buckets, points, asof_date):
    """Every output row for one route token, ordered along the route (county
    groups by odometer — ordering only, never a value — then PM)."""
    keys = [k for k in buckets if k[0] == route]
    cps = sorted({(k[2], k[3]) for k in keys})
    county_od = {}
    for county, prefix in cps:
        ods = [s.od
               for align in ("R", "L")
               for s in buckets.get((route, align, county, prefix),
                                    {}).get("HG", [])
               if s.od is not None]
        cur = county_od.get(county)
        if ods:
            county_od[county] = min(min(ods), cur) if cur is not None else min(ods)
    order = {c: i for i, c in enumerate(
        sorted(county_od, key=lambda c: county_od[c]))}

    recs = []
    for county, prefix in cps:
        rbucket = buckets.get((route, "R", county, prefix), {})
        lbucket = buckets.get((route, "L", county, prefix), {})
        pts = points.get((route, county, prefix), {})
        cp_rows = _build_cp(route, county, prefix, rbucket, lbucket, pts,
                            asof_date)
        cp_rows.sort(key=lambda r: (r["b"], {"base": 0, "R": 1,
                                             "L": 2}[r["kind"]]))
        _apply_adt(cp_rows)
        recs.extend(cp_rows)
    recs.sort(key=lambda r: (order.get(r["county"], 99),
                             r["od"] if r["od"] is not None else 1e18,
                             r["b"]))
    _finalize_route(recs)
    return [r["row"] for r in recs]


# --------------------------------------------------------------------------- #
# provenance + workbook
# --------------------------------------------------------------------------- #
def _provenance_rows(index):
    """One row per THY column — the audit record the owner asked for ('we
    will eventually want to know which layer each column comes from'):
    clean_highway_columns.PROVENANCE plus each named layer's FeatureServer
    `Data Source` from the library's 00_INDEX."""
    out = []
    for name in THY_HEADER:
        tier, layer, column, note = chc.PROVENANCE[name]
        urls = [entry["source"] for lname, entry in index.items()
                if lname and lname in layer]
        out.append([name, tier, layer, column, note,
                    " ; ".join(u for u in urls if u)])
    return out


def _write_workbook(out_path, rows, index, asof_date, lib_root, warnings):
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(SHEET_NAME)
    ws.append(THY_HEADER)
    for row in rows:
        ws.append(row)
    prov = wb.create_sheet("Provenance")
    prov.append(["THY column", "Tier", "Source layer", "Source column", "Note",
                 "FeatureServer Data Source (00_INDEX)"])
    for row in _provenance_rows(index):
        prov.append(row)
    marker = wb.create_sheet(MARKER_SHEET)
    marker.append(["Build version", BUILD_VERSION])
    marker.append(["As-of date", asof_date.isoformat()])
    marker.append(["Layer library", str(lib_root)])
    for w in warnings:
        marker.append(["Warning", w])
    tmp = out_path.with_name(out_path.name + ".tmp")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(tmp)
        os.replace(tmp, out_path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError as e:      # silent-ok: the temp is orphaned only after a failed save; the error path already reports that failure
                log.warning("could not remove %s (%s: %s)", tmp.name,
                            type(e).__name__, e)


def resolve_default_asof():
    """The fair default as-of date: the staged TSN CA HIGHWAYS extract's own
    THY_EXTRACT_DATE (read from the raw library slot), so both sides describe
    the same day. Raises ValueError when the TSN raw isn't staged."""
    import tsn_library
    from openpyxl import load_workbook

    raw_root = Path(tsn_library.raw_dir("clean_highway"))
    candidates = [p for p in sorted(raw_root.glob("*.xlsx"))
                  if p.is_file() and not p.name.startswith("~$")]
    if len(candidates) != 1:
        raise ValueError(
            "Pass an as-of date, or stage exactly one TSN CA HIGHWAYS extract "
            f"in the TSN library ({raw_root}) so the build can match its "
            "extract date.")
    wb = load_workbook(candidates[0], read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it, ())]
        if "THY_EXTRACT_DATE" not in header:
            raise ValueError(
                f"{candidates[0].name} has no THY_EXTRACT_DATE column — is it "
                "the TSN CA HIGHWAYS extract?")
        i = header.index("THY_EXTRACT_DATE")
        first = next(it, None)
        serial = crl.to_serial(first[i]) if first and i < len(first) else None
        if serial is None:
            raise ValueError(
                f"{candidates[0].name} carries no readable THY_EXTRACT_DATE — "
                "pass an as-of date explicitly.")
        return crl.serial_to_date(serial)
    finally:
        wb.close()


def _resolve_asof(asof):
    if asof is None:
        return resolve_default_asof()
    serial = crl.to_serial(asof)
    if serial is None:
        raise ValueError(f"Not a usable as-of date: {asof!r} (use YYYY-MM-DD).")
    return crl.serial_to_date(serial)


# --------------------------------------------------------------------------- #
# the public consolidate()
# --------------------------------------------------------------------------- #
def consolidate(events=None, confirm_overwrite=None, day=None, *, asof=None,
                lib_root=None, out_path=None, routes=None):
    """Build the CA HIGHWAYS clean-road workbook from the ArcGIS layer library.

    `asof` is the reconstruction date (date / ISO text / Excel serial); None
    resolves it from the staged TSN clean-road extract so the comparison is
    same-dated by default. `routes` (a set of route tokens) restricts the
    build — a development/verification convenience, never exposed in the UI."""
    del day     # ConsolidateWorker passes it; the library build has no run days
    events = events or Events()
    try:
        return _consolidate(events, confirm_overwrite, asof=asof,
                            lib_root=lib_root, out_path=out_path, routes=routes)
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))


def _consolidate(events, confirm_overwrite, *, asof, lib_root, out_path, routes):
    lib_root = Path(lib_root) if lib_root else crl.root()
    out_path = Path(out_path) if out_path else OUT_PATH
    lib = crl.inventory(lib_root)
    needed_missing = [n for n in HIGHWAY_LAYERS if n not in lib["present"]]
    if needed_missing:
        return ConsolidateResult(
            status="error",
            message=("The ArcGIS layer library is missing the highway "
                     "layer(s): " + ", ".join(needed_missing)
                     + f"\n\nDrop the per-layer .xlsx exports into:\n{lib_root}"))
    index = crl.read_index(lib_root)
    asof_date = _resolve_asof(asof)
    asof_serial = crl.to_serial(asof_date)

    if out_path.exists() and confirm_overwrite is not None \
            and not confirm_overwrite(out_path):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.",
                                 completion=outcome.CANCELLED)

    events.on_log("=" * 60)
    events.on_log(f"Clean Road Highway build — ArcGIS layers as of "
                  f"{asof_date.isoformat()}")
    events.on_log("=" * 60)

    buckets = defaultdict(lambda: defaultdict(list))
    points = defaultdict(dict)
    parked, warnings = [], []
    for tag in SPAN_LAYERS:
        if events.is_cancelled():
            return _cancelled(events)
        _read_span_layer(lib, index, tag, asof=asof_serial, events=events,
                         buckets=buckets, parked=parked)
    for tag in POINT_LAYERS:
        if events.is_cancelled():
            return _cancelled(events)
        _read_point_layer(lib, index, tag, asof=asof_serial, events=events,
                          points=points)
    _split_parked(parked, buckets, warnings)
    for w in warnings:
        events.on_log(f"  note: {w}")

    all_routes = sorted({k[0] for k in buckets if k[0]})
    if routes is not None:
        all_routes = [r for r in all_routes if r in set(routes)]
    if not all_routes:
        return ConsolidateResult(
            status="error",
            message="No as-of route coverage found in the layer library — "
                    "check the exports and the as-of date.")

    rows = []
    for i, route in enumerate(all_routes, start=1):
        if events.is_cancelled():
            return _cancelled(events)
        rows.extend(_build_route(route, buckets, points, asof_date))
        if i % 25 == 0 or i == len(all_routes):
            events.on_log(f"  built {i}/{len(all_routes)} routes "
                          f"({len(rows):,} rows)")

    _write_workbook(out_path, rows, index, asof_date, lib_root, warnings)
    result = ConsolidateResult(
        status="ok",
        message=(f"Built {len(rows):,} clean-road highway rows "
                 f"({len(all_routes)} routes) as of {asof_date.isoformat()}."
                 + (f" {len(warnings)} span(s) could not be placed — see the "
                    f"{MARKER_SHEET} sheet." if warnings else "")),
        output_path=str(out_path),
        summary_lines=[f"Clean Road Highway (ArcGIS): {len(rows):,} rows, "
                       f"{len(all_routes)} routes -> {out_path.name}"],
        completion=outcome.PARTIAL if warnings else outcome.COMPLETE,
        skipped_inputs=len(warnings), failed_inputs=0)
    if not consolidation_meta.write_outcome(
            out_path, result,
            extra={"clean_road_build": {
                "asof": asof_date.isoformat(),
                "build_version": BUILD_VERSION,
                "routes": len(all_routes),
                "rows": len(rows),
                "warnings": warnings,
            }}):
        return ConsolidateResult(
            status="error",
            message="The build finished but its outcome sidecar could not be "
                    "published — rebuild before comparing.")
    # The sidecar above is authoritative (it carries the build facts); the
    # generic worker-boundary write must not overwrite it.
    result.sidecar_published = True
    events.on_log(result.message)
    return result
