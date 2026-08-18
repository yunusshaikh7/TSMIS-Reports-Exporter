"""Consolidate Highway Summary per-route XLSX into one workbook.

Each per-route export (sheet "Highway Summary") is a MILES-measured statistics
document: a "TOTAL MILES SELECTED" scalar followed by 10 category sections, each
a `Code | Miles` block. This consolidator reads every route through the ONE
strict reader in `highway_summary_columns.values_from_rows` (the same reader the
cross-environment loader uses, so the two paths cannot drift — the CMP-AUD-018
rule) and writes:

  * "Highway Summary"  one row per route — Route, Total Miles, and one column per
                       category (its compare key as the header).
  * "Combined"         the familiar section-grouped statewide rollup (the source
                       arrangement) with the summed mileage.

Mileage is summed in exact integer THOUSANDTHS and converted once for display, so
a statewide rollup is exact rather than a drifting float accumulation.

Console-free (Events sink + ConsolidateResult; no print/input/sys.exit),
mirroring consolidate_intersection_summary. Importable:
consolidate(events, confirm_overwrite, day=None) returns a ConsolidateResult so
the GUI can drive it; the console UX lives in cli.run_consolidate_cli.
"""
import re
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import artifact_store
import consolidation_meta
import highway_summary_columns as hsc
import outcome
from compare_core import is_formula_injection
from events import ConsolidateResult, Events
from paths import (OUTPUT_ROOT, latest_output_day, output_day_dir,
                   stamped_consolidated_filename)

SUBDIR = "highway_summary"
FILENAME = "highway_summary_consolidated.xlsx"
SHEET_NAME = hsc.SHEET_NAME                  # per-route source sheet AND output sheet
COMBINED_SHEET = "Combined"
REPORT_NAME = "Highway Summary"

# Input format badge in the Consolidate list.
INPUT_FMT = "Excel"

# Legacy flat-layout locations (pre-dated exports); still used when no dated
# output/<YYYY-MM-DD>/ folders exist, so old exports stay consolidatable.
INPUT_DIR = OUTPUT_ROOT / SUBDIR
OUT_DIR = OUTPUT_ROOT / "consolidated"
OUT_PATH = OUT_DIR / FILENAME

_ROUTE_FROM_NAME = re.compile(r"_route_(\w+)\.xlsx$", re.IGNORECASE)


def input_dir_for(day):
    """Per-route exports for `day` (a run-folder name); None = the legacy flat layout."""
    return (output_day_dir(day) / SUBDIR) if day else INPUT_DIR


def out_path_for(day):
    if not day:
        return OUT_PATH
    return output_day_dir(day) / "consolidated" / stamped_consolidated_filename(FILENAME, day)


# --------------------------------------------------------------------------- #
# parse one per-route XLSX -> (route, {slug: milli}, total_milli)
# --------------------------------------------------------------------------- #
def parse_route(path):
    """Read one per-route Highway Summary workbook. Returns
    (route, values{slug: thousandths}, total_thousandths).

    The export sheet carries no route identity of its own (unlike Intersection
    Summary's 'Route: NNN' header), so the route comes from the filename's
    `_route_<token>` — the end-anchored contract every export writes."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = [(r[0] if len(r) > 0 else None, r[1] if len(r) > 1 else None)
                for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    total, values = hsc.values_from_rows(rows, source=Path(path).name)
    m = _ROUTE_FROM_NAME.search(Path(path).name)
    route = m.group(1) if m else Path(path).stem
    return route, values, total


def record_has_data(rec):
    """True if a parsed route carries real mileage (not just a route id).

    ANY non-zero measure counts, not just the total: on the 2026-08-17 census
    four routes (010S/014U/015S/178S) report TOTAL MILES SELECTED = 0 while
    carrying real NON-ADD mileage (0.692–18.785 miles) and 10–102 Highway Detail
    records apiece. TOTAL MILES SELECTED tabulates ADD mileage only, so keying
    'has data' on it alone would drop four real routes from the statewide table
    and pin every run to PARTIAL."""
    return bool(rec.get("total")) or any(rec.get("values", {}).values())


def record_problem(values, total, *, source):
    """The strict parser-integrity problem for one parsed route, or None when it
    looks structurally sound. Shared by the consolidation AND the cross-
    environment loader (CMP-AUD-018) so a misread record can't slip through one
    path while the other gates it. The skeleton itself is already enforced by
    `values_from_rows`; this adds the partition contract."""
    return hsc.partition_problem(total, values, source=source)


# --------------------------------------------------------------------------- #
# records -> workbook
# --------------------------------------------------------------------------- #
_TITLE_FILL = "1F3864"
_SECTION_FILL = "0070C0"
_HEADER_FILL = "305496"
_MILES_FMT = "#,##0.000"


def _build_combined(wb, statewide, total, notes):
    """A familiar section-grouped statewide rollup sheet (the source arrangement)."""
    ws = wb.create_sheet(COMBINED_SHEET, 0)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_title = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    f_sec = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    f_body = Font(name="Arial", size=10)
    f_total = Font(name="Arial", bold=True, size=12)
    f_note = Font(name="Arial", size=9, italic=True, color="595959")
    right = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws["A1"] = "All Routes Combined — TSAR Highway Summary"
    ws.merge_cells("A1:B1")
    ws["A1"].font = f_title
    ws["A1"].fill = PatternFill("solid", start_color=_TITLE_FILL)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = hsc.TOTAL_LABEL
    ws["A2"].font = f_total
    ws["B2"] = hsc.miles(total)
    ws["B2"].font = f_total
    ws["B2"].alignment = right
    ws["B2"].number_format = _MILES_FMT

    r = 4
    for section in hsc.SECTIONS:
        ws.cell(r, 1, section.name).font = f_sec
        ws.cell(r, 1).fill = PatternFill("solid", start_color=_SECTION_FILL)
        ws.cell(r, 2, "").fill = PatternFill("solid", start_color=_SECTION_FILL)
        r += 1
        for cat in section.cats:
            ws.cell(r, 1, cat.label).font = f_body
            ws.cell(r, 1).border = border
            c = ws.cell(r, 2, hsc.miles(statewide.get(cat.slug, 0)))
            c.font, c.alignment, c.border = f_body, right, border
            c.number_format = _MILES_FMT
            r += 1
        r += 1

    if notes:
        ws.cell(r, 1, "Sections that tabulate less than the total (the site's own "
                      "classification; the residual is shown, never redistributed):"
                ).font = f_note
        r += 1
        for n in notes:
            ws.cell(r, 1, n).font = f_note
            r += 1
    wb.active = wb.index(ws)


def build_workbook(records, out_path, statewide, total_all, notes,
                   proceed=None, commit_guard=None):
    """Per-route sheet (Route, Total Miles, one column per category) + Combined.
    `proceed` (P12) is the pre-replace overwrite gate atomic_save_if evaluates
    JUST BEFORE the os.replace; returns True iff committed."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color=_HEADER_FILL)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, label in enumerate(hsc.HEADER, start=1):
        cell = ws.cell(1, ci, label)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align

    for r, rec in enumerate(records, start=2):
        ws.cell(r, 1, rec["route"]).alignment = Alignment(horizontal="left")
        if is_formula_injection(rec["route"]):
            ws.cell(r, 1).data_type = "s"
        c = ws.cell(r, 2, hsc.miles(rec["total"]))
        c.alignment, c.number_format = Alignment(horizontal="right"), _MILES_FMT
        for ci, cat in enumerate(hsc.CATS, start=3):
            c = ws.cell(r, ci, hsc.miles(rec["values"].get(cat.slug, 0)))
            c.alignment, c.number_format = Alignment(horizontal="right"), _MILES_FMT

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12

    _build_combined(wb, statewide, total_all, notes)
    if not consolidation_meta.guard_allows(commit_guard, out_path):
        return False
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not consolidation_meta.guard_allows(commit_guard, out_path):
        return False
    # F9 temp + os.replace + the P12 TOCTOU gate at the replace.
    return artifact_store.atomic_save_if(wb, out_path, proceed or (lambda: True))


# --------------------------------------------------------------------------- #
# entry point
# --------------------------------------------------------------------------- #
def consolidate(events=None, confirm_overwrite=None, day=None,
                input_dir=None, out_path=None, commit_guard=None):
    """Parse every per-route Highway Summary XLSX into one workbook.
    Console-free; honors cancel; returns a ConsolidateResult."""
    events = events or Events()
    if not _DEPS_OK:
        return ConsolidateResult(status="error",
                                 message="Required components are missing (openpyxl).")
    confirm = confirm_overwrite or (lambda _p: True)
    day = day or latest_output_day()
    input_dir = input_dir or input_dir_for(day)
    out_path = out_path or out_path_for(day)

    if not input_dir.exists():
        return ConsolidateResult(
            status="error",
            message=(f"The {REPORT_NAME} output folder doesn't exist yet:\n{input_dir}\n\n"
                     f"Export the {REPORT_NAME} report first, then consolidate."))
    files = [f for f in sorted(input_dir.glob("*.xlsx")) if not f.name.startswith("~$")]
    if not files:
        return ConsolidateResult(
            status="error",
            message=(f"No {REPORT_NAME} files were found in:\n{input_dir}\n\n"
                     f"Export the {REPORT_NAME} report first, then consolidate."))
    existed_at_confirm = out_path.exists()
    if existed_at_confirm and not confirm(out_path):
        return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"TSAR Highway Summary Consolidation - {len(files)} file(s)")
    events.on_log("=" * 60)

    records, failed, blank = [], [], []
    for i, p in enumerate(files, 1):
        if events.is_cancelled():
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i:>3}/{len(files)}] {p.name}"
        try:
            route, values, total = parse_route(str(p))
        except Exception as e:
            events.on_log(f"{prefix} FAILED ({type(e).__name__}): {e}")
            failed.append(p.name)
            continue
        # CMP-AUD-183: a statewide aggregation needs a sound route identity per
        # row — a blank/malformed identity can't be attributed and must not sum.
        route = str(route).strip() if route is not None else ""
        if not route or not route.isalnum() or len(route) > 8:
            events.on_log(f"{prefix} FAILED (no usable route identity: {route!r})")
            failed.append(p.name)
            continue
        rec = {"route": route, "values": values, "total": total, "file": p.name}
        problem = record_problem(values, total, source=p.name)
        if problem:
            events.on_log(f"{prefix} FAILED (parse integrity): {problem}")
            failed.append(p.name)
        elif record_has_data(rec):
            records.append(rec)
            events.on_log(f"{prefix} parsed (route {route}, "
                          f"{hsc.miles(total):,.3f} miles)")
            for note in hsc.partition_notes(total, values):
                events.on_log(f"          route {route} {note}")
        else:
            events.on_log(f"{prefix} skipped: no mileage")
            blank.append(p.name)

    # CMP-AUD-183: two inputs claiming the SAME route identity make the statewide
    # table ambiguous (identical duplicates double-count; conflicting ones can't
    # be arbitrated here) — exclude every claimant loudly, so the run reports
    # PARTIAL instead of silently aggregating a corrupt universe.
    by_route = {}
    for rec in records:
        by_route.setdefault(rec["route"], []).append(rec)
    duplicated = {route: recs for route, recs in by_route.items() if len(recs) > 1}
    if duplicated:
        for route, recs in sorted(duplicated.items()):
            names = ", ".join(rec["file"] for rec in recs)
            events.on_log(f"FAILED (duplicate route {route}): {len(recs)} files "
                          f"claim the same route — {names}")
            failed.extend(rec["file"] for rec in recs)
        records = [rec for rec in records if rec["route"] not in duplicated]

    if not records:
        return ConsolidateResult(
            status="error",
            message=(f"None of the {len(files)} {REPORT_NAME} file(s) yielded data "
                     f"({len(failed)} failed, {len(blank)} empty). Nothing was written."))

    # Statewide rollup in exact thousandths, converted once for display.
    statewide = {c.slug: 0 for c in hsc.CATS}
    total_all = 0
    for rec in records:
        total_all += rec["total"]
        for slug, v in rec["values"].items():
            statewide[slug] += v
    notes = hsc.partition_notes(total_all, statewide)

    events.on_log("")
    events.on_log("Writing consolidated workbook...")
    try:
        # P12 TOCTOU: the overwrite gate is INSIDE build_workbook, at the os.replace
        # (atomic_save_if) — a destination that appears during the BUILD is caught.
        committed = build_workbook(
            records, out_path, statewide, total_all, notes,
            proceed=lambda: (consolidation_meta.guard_allows(commit_guard, out_path)
                             and artifact_store.confirm_late_overwrite(
                                 out_path, existed_at_confirm, confirm)),
            commit_guard=commit_guard)
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not save {out_path.name}.\n\n"
                     "The file is probably open in Excel. Close it and try again."))
    if not committed:
        return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    incomplete = bool(failed or blank)
    summary_lines = []
    if incomplete:
        summary_lines.append(
            f"⚠ INCOMPLETE — {len(failed) + len(blank)} file(s) left OUT "
            f"({len(failed)} failed, {len(blank)} empty). Re-export before relying on it.")
    summary_lines += [
        f"Parsed:      {len(records)}",
        f"Total miles: {hsc.miles(total_all):,.3f}",
        f"Failed:      {len(failed)} {failed if failed else ''}",
        f"Empty:       {len(blank)} {blank if blank else ''}",
        f"Output file: {out_path}",
    ]
    # CMP-AUD-183: persist the ordered route identities the workbook was built
    # from, so the comparison loader can reconcile the aggregated universe
    # against the producer's own census.
    return ConsolidateResult(status="ok", output_path=str(out_path),
                             summary_lines=summary_lines,
                             completion=outcome.PARTIAL if incomplete else outcome.COMPLETE,
                             skipped_inputs=len(blank), failed_inputs=len(failed),
                             producer_extra={
                                 "route_census": [rec["route"] for rec in records]})


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
