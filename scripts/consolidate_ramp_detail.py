"""Consolidate TSAR: Ramp Detail XLSX files into a single workbook.

Reads every XLSX in   output/<date>/ramp_detail/   (newest day by default)
Writes one workbook in output/<date>/consolidated/tsar_ramp_detail_consolidated.xlsx
with a leading "Route" column added so rows from different routes are
distinguishable in the combined file.

Thin wrapper over consolidate_xlsx_base, which is shared with Highway Sequence
and Highway Log -- all three are "one sheet, header row + data rows" exports that
differ only by input folder, sheet name, and output name. (The Ramp Summary
consolidator stays standalone because it parses PDFs, not XLSX.) HF-04 adds one
Ramp-Detail-specific step: the produced workbook's header is verified against
the comparator's own consumability gate, so a consolidation whose output no
comparison would accept reports an error instead of ok.

Importable (Phase 3b): consolidate(events, confirm_overwrite, day=None) returns
a ConsolidateResult and never prints/prompts/exits, so the GUI can drive it. The
console UX lives in cli.run_consolidate_cli, used by the __main__ entry (and
therefore by "4. consolidate (combine reports).bat").
"""
import outcome
from consolidate_xlsx_base import consolidate_xlsx
from events import ConsolidateResult
from paths import (OUTPUT_ROOT, latest_output_day, output_day_dir,
                   stamped_consolidated_filename)

SUBDIR = "ramp_detail"
FILENAME = "tsar_ramp_detail_consolidated.xlsx"

# Legacy flat-layout locations (pre-dated exports); still used when no dated
# output/<YYYY-MM-DD>/ folders exist, so old exports stay consolidatable.
INPUT_DIR = OUTPUT_ROOT / SUBDIR
OUT_DIR = OUTPUT_ROOT / "consolidated"
OUT_PATH = OUT_DIR / FILENAME

# Sheet name produced by the TSMIS export — must match exactly.
SHEET_NAME = "TSAR - Ramp Detail"

# Friendly report name for user-facing messages (shown in both the GUI and the
# console, so keep it UI-neutral -- no ".bat" / "menu option" wording).
REPORT_NAME = "Ramp Detail"


def input_dir_for(day):
    """Per-route exports for `day` (a run-folder name); None = the legacy flat layout."""
    return (output_day_dir(day) / SUBDIR) if day else INPUT_DIR


def out_path_for(day):
    """Consolidated workbook destination for `day` (a run-folder name); None = the
    legacy location. The filename carries the run's date + source/environment (A1)
    so a copy lifted out of its folder keeps its provenance."""
    if not day:
        return OUT_PATH
    return output_day_dir(day) / "consolidated" / stamped_consolidated_filename(FILENAME, day)


def _consumability_downgrade(result):
    """HF-04 / PCOA-FINAL-001: a consolidation no comparator accepts must not
    report ok. The shared consolidator locks whatever header the first readable
    export carries, so a NEW site layout consolidates "successfully" into a
    workbook every comparison then refuses — the user gets a green step followed
    by refusals that blame the wrong thing. Verify the PRODUCED workbook's
    header against the comparator's own consumability gate
    (compare_ramp_detail_tsn.consolidated_header_ok — the same predicate
    _load_tsmis enforces, so "consolidation ok" and "a comparator will read it"
    cannot drift apart) and downgrade the result when it fails. The combined
    file is left on disk for inspection but the result is an error with
    completion=failed, so nothing promotes, caches, or compares it as good."""
    if result.status != "ok" or not result.output_path:
        return result
    import compare_ramp_detail_tsn as _rd    # lazy: keep module import light
    from openpyxl import load_workbook       # present iff consolidation ran
    try:
        wb = load_workbook(result.output_path, read_only=True, data_only=True)
        try:
            ws = wb[SHEET_NAME]
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header = [("" if c is None else str(c).strip()) for c in row]
        finally:
            wb.close()
    except Exception as e:  # noqa: BLE001 — any unreadable result is unverifiable
        return ConsolidateResult(
            status="error", completion=outcome.FAILED,
            message=(f"The combined {REPORT_NAME} workbook was written but its "
                     f"column layout could not be verified "
                     f"({type(e).__name__}) — treat it as unusable and "
                     f"re-consolidate.\nFile: {result.output_path}"),
            skipped_inputs=result.skipped_inputs,
            failed_inputs=result.failed_inputs)
    if _rd.consolidated_header_ok(header):
        return result
    return ConsolidateResult(
        status="error", completion=outcome.FAILED,
        message=(f"The per-route {REPORT_NAME} exports use a column layout "
                 "this app version does not support, so the combined workbook "
                 "cannot be used by any comparison. It was kept for "
                 f"inspection at:\n{result.output_path}\n\n"
                 "This app supports the classic and the July-2026 site export "
                 "layouts. If the site's export format has changed again, "
                 "this app needs an update before Ramp Detail can be "
                 "consolidated and compared."),
        skipped_inputs=result.skipped_inputs,
        failed_inputs=result.failed_inputs)


def consolidate(events=None, confirm_overwrite=None, day=None,
                input_dir=None, out_path=None, commit_guard=None):
    """Combine every per-route Ramp Detail XLSX into one workbook.

    `day` picks which export run folder ("<YYYY-MM-DD> <src>-<env>") to read; None means
    the newest run folder, falling back to the legacy flat layout when no run
    folders exist yet."""
    day = day or latest_output_day()
    result = consolidate_xlsx(
        input_dir=input_dir or input_dir_for(day),
        out_path=out_path or out_path_for(day),
        sheet_name=SHEET_NAME, report_name=REPORT_NAME,
        title="TSAR Ramp Detail Consolidation",
        events=events, confirm_overwrite=confirm_overwrite,
        commit_guard=commit_guard,
    )
    return _consumability_downgrade(result)


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
