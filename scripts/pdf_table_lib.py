"""Shared machinery for the table-PDF parsers/consolidators (v0.19.0 R2).

Before this module, the three PDF consolidators (TSMIS Highway Log, TSMIS
Intersection Detail, TSN Highway Log) + the TSN Highway Sequence loader each
carried verbatim copies of the same building blocks: the y-cluster line builder
(x4), the char->column assigner (x2), the median/contiguous column-window math
(x2), the route-token normalizer (x4, one divergent), the TSMIS-format
route-workbook writer (x3), and the ~110-line per-PDF convert loop (x2). This
module is those blocks, parameterized; each consolidator keeps ONLY its layout
knowledge (band collection, row classification/assembly) — which is what makes
a NEW PDF-sourced report (Highway Detail/Summary) a recipe instead of a port.

`norm_route` is THE canonical route-token normalizer, reconciling the four
copies deliberately:

  * the two TSMIS-PDF copies and the TSN Highway Sequence copy were already
    semantically identical (regex: int-collapse the digits, zero-pad to 3,
    upper-case the suffix) — adopting this changes nothing for them;
  * the TSN Highway Log copy used `zfill(3)`, which (a) never padded a SHORT
    SUFFIXED token ('5S' stayed '5S' while TSMIS prints '005S' — a latent
    row-misalignment) and (b) kept over-padded digits ('0001' stayed '0001').
    The reconciled form fixes both. Real district PDFs print 1-3 digit tokens
    (suffix included in the 'nnnX' form), where the two agree — but because the
    normalized route KEYS the stored TSN library, the highway_log TSN entry's
    `normalization_version` is bumped with this change (D2 auto-rebuild).

Console-free; pdfplumber/openpyxl objects come from the callers (which gate on
their own _DEPS_OK), so this module keeps the same lazy posture and never
imports pdfplumber itself.
"""
import os
import re
import uuid

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from compare_core import is_formula_injection   # shared formula-injection guard
from consolidate_xlsx_base import consolidate_xlsx
import consolidation_meta
from events import ConsolidateResult, Events
import owned_dir


# --------------------------------------------------------------------------- #
# tokens
# --------------------------------------------------------------------------- #
def norm_route(token):
    """'6' -> '006' (TSMIS zero-pads to 3 digits); suffixed routes ('101U',
    '005S') keep their letter and are upper-cased; anything else upper-cased
    as-is. See the module docstring for the 4-copy reconciliation."""
    m = re.fullmatch(r"(\d+)([A-Za-z]?)", token)
    if not m:
        return token.upper()
    return f"{int(m.group(1)):03d}{m.group(2).upper()}"


def unexpected_pm_tokens(prefix, suffix, *, prefix_set, suffix_set=frozenset()):
    """CMP-AUD-063: the post-mile CODE tokens outside the accepted, versioned
    vocabulary, as a list of ``(window, token)`` pairs — e.g. ``[("prefix",
    "Q")]``. A window's text is UNEXPECTED when it is non-empty and not an EXACT
    member of its accepted set, so a lowercase ``'c'``, a joined ``'CE'``, a
    multi-letter ``'QQ'``, or an unknown ``'Q'`` each qualify, while a blank
    window (the common case) and an exact legend code never do. ``suffix`` /
    ``suffix_set`` default to ``""`` / the empty set for a report with no suffix
    column (Ramp Detail), so any suffix text there is unexpected by construction.

    The comparison is deliberately exact — no normalization — because the
    accepted vocabulary IS the source of truth: anything else alters the
    canonical post-mile key and marks the source as suspect, so the producer
    escalates to at least PARTIAL rather than certifying it complete. The two
    parsers own their censused `PREFIX_SET`/`SUFFIX_SET` + `PM_VOCAB_VERSION`;
    this shared helper owns only the membership rule."""
    bad = []
    if prefix and prefix not in prefix_set:
        bad.append(("prefix", prefix))
    if suffix and suffix not in suffix_set:
        bad.append(("suffix", suffix))
    return bad


class RouteIdentityError(ValueError):
    """CMP-AUD-049 (evidence half): a per-route PDF's own claims failed to
    confirm the expected route — the document must not be captioned/verified
    as that route."""


def require_document_route(pdf_name, expected_route, doc_routes, *, claim_desc):
    """Raise RouteIdentityError unless the document's own claims confirm
    `expected_route` (the evidence adapters' twin of
    reconcile_route_identity — same identity rule, exception-shaped for the
    locate paths, which have no failed-input ledger to write to)."""
    claims = sorted(set(doc_routes))
    if expected_route and claims == [expected_route]:
        return
    if not claims:
        raise RouteIdentityError(
            f"{pdf_name}: the document does not identify its route "
            f"({claim_desc} was not found), so it cannot verify route "
            f"{expected_route or '?'} evidence")
    raise RouteIdentityError(
        f"{pdf_name}: the document claims route "
        f"{', '.join(claims)} ({claim_desc}) but route "
        f"{expected_route or '?'} was expected — was the file renamed or "
        "mixed up?")


def reconcile_route_identity(pdf_name, name_route, doc_routes, events, ctx,
                             *, claim_desc):
    """CMP-AUD-049 (converter half): the DOCUMENT's own route claim is the
    authoritative identity of a per-route PDF; a filename token merely
    corroborates it. Returns the confirmed normalized route, or None after
    logging why and naming the file in ctx["failed"] (each report's finalize
    escalates PARTIAL, so a mis-named or unidentifiable input is never
    silently absorbed under the wrong route).

    `doc_routes`: every distinct normalized route the document claims for
    itself — its page banners / cover line / per-record Location cells
    (`claim_desc` names the family's source). Empty means the document never
    identified itself; more than one means it contradicts itself; a filename
    token that disagrees means the file was probably renamed or mixed up.
    All three refuse loudly rather than guessing.
    """
    claims = sorted(set(doc_routes))
    if not claims:
        events.on_log(
            f"  {pdf_name}: the document does not identify its route "
            f"({claim_desc} was not found) — cannot confirm which route this "
            "PDF belongs to; skipping. Re-export that route's PDF.")
        ctx["failed"].append(pdf_name)
        return None
    if len(claims) > 1:
        events.on_log(
            f"  {pdf_name}: the document claims more than one route for "
            f"itself ({', '.join(claims)} via {claim_desc}) — skipping. "
            "Re-export that route's PDF.")
        ctx["failed"].append(pdf_name)
        return None
    doc_route = claims[0]
    if name_route and name_route != doc_route:
        events.on_log(
            f"  {pdf_name}: the filename says route {name_route} but the "
            f"document says route {doc_route} ({claim_desc}) — the file was "
            "probably renamed or mixed up; skipping. Re-export that route's "
            "PDF or restore its original name.")
        ctx["failed"].append(pdf_name)
        return None
    return doc_route


def median(values):
    """The parsers' median: upper-middle element (robust to a stray rect)."""
    s = sorted(values)
    return s[len(s) // 2]


# --------------------------------------------------------------------------- #
# line clustering + column assignment
# --------------------------------------------------------------------------- #
def cluster_by_top(items, y_tolerance):
    """Cluster x0/top-carrying items (chars or words) into logical lines,
    tolerating the small baseline jitter of a wrapped row. Returns
    [(anchor_top, [item, ... x0-sorted]), ...] in reading order."""
    clusters = []                     # [(anchor_top, [item, ...]), ...]
    for it in sorted(items, key=lambda i: (i["top"], i["x0"])):
        if clusters and abs(it["top"] - clusters[-1][0]) <= y_tolerance:
            clusters[-1][1].append(it)
        else:
            clusters.append((it["top"], [it]))
    return [(top, sorted(cs, key=lambda i: i["x0"])) for top, cs in clusters]


def words_from_chars(chars, word_gap):
    """Build word tokens from one x0-sorted char line: chars closer than
    `word_gap` fuse into one token."""
    words = []
    for c in chars:
        if words and c["x0"] - words[-1]["x1"] < word_gap:
            words[-1]["text"] += c["text"]
            words[-1]["x1"] = c["x1"]
        else:
            words.append({"text": c["text"], "x0": c["x0"], "x1": c["x1"]})
    return words


def char_lines(page, y_tolerance, word_gap):
    """Cluster the page's non-space characters into logical lines. Each line
    yields its word tokens (for classifying the line) AND the raw characters
    (for column parsing — adjacent columns can sit closer than any word
    tolerance, so values are assigned char by char, never word by word)."""
    chars = [c for c in page.chars if c["text"].strip()]
    return [(top, words_from_chars(cs, word_gap), cs)
            for top, cs in cluster_by_top(chars, y_tolerance)]


def assign_columns(chars, windows, word_gap):
    """Map each character of a data line to its column by horizontal center.
    Characters of one column abut (~0pt apart); a gap >= `word_gap` inside the
    same column means two tokens, kept apart with a space. Values are returned
    RAW (unstripped) — callers strip to taste."""
    vals = ["" for _ in windows]
    last_x1 = [None] * len(windows)
    for c in chars:                   # x-sorted by the clusterer
        center = (c["x0"] + c["x1"]) / 2
        for i, (lo, hi) in enumerate(windows):
            if lo <= center < hi:
                if vals[i] and c["x0"] - last_x1[i] >= word_gap:
                    vals[i] += " "
                vals[i] += c["text"]
                last_x1[i] = c["x1"]
                break
    return vals


def contiguous_windows(edges_lo, edges_hi):
    """Column x-windows from per-column median edges, made CONTIGUOUS (each
    boundary is the midpoint between adjacent cells; the first/last extend to
    ±infinity) so no data character can fall between two cells and be
    silently dropped."""
    n = len(edges_lo)
    windows = []
    for i in range(n):
        lo = float("-inf") if i == 0 else (edges_hi[i - 1] + edges_lo[i]) / 2
        hi = float("inf") if i == n - 1 else (edges_hi[i] + edges_lo[i + 1]) / 2
        windows.append((lo, hi))
    return windows


def carried_line_crossings(chars, windows, word_gap):
    """How badly a data line disagrees with CARRIED-FORWARD column windows:
    the number of intra-token window splits — consecutive characters closer
    than `word_gap` (one printed token) whose centers fall in different
    windows. A page that shares the carried table's layout scores 0 (a
    printed cell never straddles a column boundary); a drifted or
    foreign-table geometry cuts through tokens and scores immediately.
    Read-only, and it applies the same char-center window test as
    `assign_columns` — so a 0 score certifies that assignment placed every
    token of the line intact."""
    n = 0
    prev_win = None
    prev_x1 = None
    for c in chars:                   # x-sorted by the clusterer
        center = (c["x0"] + c["x1"]) / 2
        win = None
        for i, (lo, hi) in enumerate(windows):
            if lo <= center < hi:
                win = i
                break
        if (prev_win is not None and prev_x1 is not None
                and (c["x0"] - prev_x1) < word_gap and win != prev_win):
            n += 1
        prev_win, prev_x1 = win, c["x1"]
    return n


# --------------------------------------------------------------------------- #
# TSMIS-format per-route workbooks
# --------------------------------------------------------------------------- #
# CMP-AUD-066: every workbook this app writes FROM PDFs carries a very-hidden
# versioned marker sheet, so the PDF-vs-Excel comparison flavors can prove
# their "TSMIS (PDF)" side really is a PDF conversion and their "TSMIS (Excel)"
# side really is not. The TSN Highway Log consolidator shares
# write_route_workbook and must stay UNMARKED (it is not a TSMIS PDF
# conversion; it carries its own "TSN Normalization" marker instead) — hence
# the opt-in `pdf_source_marker=` seam rather than an unconditional stamp.
PDF_SOURCE_MARKER_SHEET = "TSMIS PDF Conversion"
PDF_SOURCE_MARKER_VERSION = 1


def write_pdf_source_marker(wb):
    """Stamp `wb` as produced by this app's TSMIS PDF conversion (066).
    append()-based so it works on BOTH ordinary and write-only workbooks
    (consolidate_xlsx builds the combined workbook write-only)."""
    ws = wb.create_sheet(PDF_SOURCE_MARKER_SHEET)
    ws.append([PDF_SOURCE_MARKER_SHEET])
    ws.append([PDF_SOURCE_MARKER_VERSION])
    ws.sheet_state = "veryHidden"


def pdf_source_marker_state(path):
    """The workbook's PDF-conversion marker: >0 = the marker's version, 0 = no
    marker sheet at all, -1 = a marker sheet exists but is malformed. Callers
    fail closed on BOTH roles: requiring the marker accepts only >0, and
    rejecting it refuses anything != 0 (a corrupted marker still says
    "PDF-sourced", it just can't certify a version)."""
    try:
        wb = load_workbook(path, read_only=True)
    except Exception:  # silent-ok: the caller's own load reports unreadable files
        return 0
    try:
        if PDF_SOURCE_MARKER_SHEET not in wb.sheetnames:
            return 0
        vals = [r[0] for r in wb[PDF_SOURCE_MARKER_SHEET].iter_rows(
            min_row=1, max_row=2, max_col=1, values_only=True)]
        if (len(vals) == 2 and vals[0] == PDF_SOURCE_MARKER_SHEET
                and isinstance(vals[1], (int, float)) and int(vals[1]) > 0):
            return int(vals[1])
        return -1
    except Exception:  # silent-ok: malformed content = present-but-unreadable
        return -1
    finally:
        wb.close()


def write_route_workbook(rows, out_path, *, sheet_name, header,
                         row_values=None, apply_tooltips=None, decorate=None,
                         pdf_source_marker=False):
    """Write one route's rows as a TSMIS-format workbook (the same sheet name +
    columns the report's Excel export uses): the standard blue header row,
    frozen panes, per-column widths (Description wide), the formula-injection
    guard on every data cell, and the report's optional header tooltips /
    workbook decoration (e.g. the Highway Log Legend sheet).
    `pdf_source_marker=True` (the five TSMIS PDF consolidators) stamps the
    CMP-AUD-066 provenance marker; the TSN consolidator leaves it off."""
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(header))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    if apply_tooltips is not None:
        apply_tooltips(ws)                   # hover any header for its meaning
    ws.freeze_panes = "A2"
    for i, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = \
            40 if name == "Description" else 12

    for row in rows:
        ws.append(row_values(row) if row_values is not None else row)
        # Neutralize any formula-looking text (e.g. a Description starting with
        # "=") so it can't execute when the workbook is opened.
        for cell in ws[ws.max_row]:
            if is_formula_injection(cell.value):
                cell.data_type = "s"
    if decorate is not None:
        decorate(wb)                         # e.g. a "Legend" tab
    if pdf_source_marker:
        write_pdf_source_marker(wb)
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# the per-PDF convert loop (the TSMIS PDF consolidators' shared driver)
# --------------------------------------------------------------------------- #
def run_pdf_conversion(*, in_dir, out, conv, deps_ok, events, confirm_overwrite,
                       report_name, banner_title, export_hint, unreadable_hint,
                       converted_prefix, convert_one, write_one, finalize,
                       consolidate_kwargs, commit_guard=None):
    """The convert-every-PDF-then-combine driver both TSMIS PDF consolidators
    wrote verbatim: deps gate -> glob + the no-inputs error -> the EARLY
    overwrite confirm (before any parsing) -> banner -> clear the stale
    converted files -> the per-PDF loop (cancel, parse via `convert_one`,
    the CMP-AUD-050 route-universe gate — a blank route identity or two PDFs
    converting to the same route REFUSES with both sources named, never
    overwriting or double-counting by file order — write via `write_one`,
    the locked-file error) -> the none-readable error -> `consolidate_xlsx`
    -> the shared summary lines + the report's `finalize` (its ⚠ notes +
    producer-owned PARTIAL escalation).

    `convert_one(p, prefix, events, ctx)` owns the per-report step: route
    resolution (CMP-AUD-049: the document's own route claim is authoritative —
    each family passes its in-document claims through
    `reconcile_route_identity`, and a missing/conflicting/filename-disagreeing
    claim is a named FAILED input), parse, its own stats in `ctx`, and its own
    skip logging (it appends skipped names to ctx["failed"]). It returns
    ("ok", route, rows), ("skip",), or ("cancelled",). `write_one(rows,
    out_file)` writes one route workbook. `finalize(result, ctx)` runs only on
    a successful combine.
    """
    events = events or Events()

    def destination_changed():
        return ConsolidateResult(
            status="error",
            message="The destination changed while converting PDFs; nothing was published.")
    if not deps_ok:
        return ConsolidateResult(
            status="error",
            message="Required components are missing (pdfplumber, openpyxl).",
        )
    confirm = confirm_overwrite or (lambda _p: True)

    # Ensure the folder exists so the error below names a real, openable path.
    try:
        in_dir.mkdir(parents=True, exist_ok=True)
    except OSError:  # silent-ok: the no-inputs error below names the path
        pass

    pdfs = sorted(in_dir.glob("*.pdf"))
    if not pdfs:
        return ConsolidateResult(
            status="error",
            message=(f"No {report_name} files were found in:\n{in_dir}\n\n"
                     f"{export_hint}"),
        )

    # Confirm overwrite *before* spending time parsing PDFs.
    existed_at_confirm = out.exists()
    if existed_at_confirm and not confirm(out):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"{banner_title} - {len(pdfs)} route PDF(s)")
    events.on_log("=" * 60)
    events.on_log("")

    # The combined workbook reflects exactly THIS run's PDFs: clear previously
    # converted files so routes removed from the input folder don't linger.
    if not consolidation_meta.guard_allows(commit_guard, conv):
        return destination_changed()
    conv.mkdir(parents=True, exist_ok=True)
    conv_identity = (owned_dir.directory_identity(conv)
                     if commit_guard is not None else None)

    def guarded(path):
        return (consolidation_meta.guard_allows(commit_guard, path)
                and (commit_guard is None
                     or (conv_identity is not None
                         and owned_dir.directory_identity(conv) == conv_identity)))

    if not guarded(conv):
        return destination_changed()
    stale = list(conv.glob(f"{converted_prefix}_*.xlsx"))
    for p in stale:
        if not guarded(p):
            return destination_changed()
        try:
            p.unlink()
        except OSError:
            return ConsolidateResult(
                status="error",
                message=(f"Could not replace {p.name}.\n\n"
                         "The file is probably open in Excel. Close it and try again."),
            )
    if stale:
        events.on_log(f"Cleared {len(stale)} previously converted file(s).")

    ctx = {"failed": [], "pdfs": len(pdfs)}
    converted = 0
    route_sources = {}               # route -> the PDF that produced it (CMP-AUD-050)
    for i, p in enumerate(pdfs, 1):
        if events.is_cancelled():
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i}/{len(pdfs)}] {p.name}"
        events.on_log(f"{prefix} parsing…")
        try:
            step = convert_one(p, prefix, events, ctx)
        except Exception as e:                       # noqa: BLE001 — one bad PDF must not sink the run
            events.on_log(f"{prefix} FAILED ({type(e).__name__}): {e}")
            ctx["failed"].append(p.name)
            continue
        if step[0] == "cancelled":
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        if step[0] == "skip":
            continue
        _tag, route, rows = step
        # CMP-AUD-050: every converted PDF must own exactly one nonblank
        # route, and two PDFs claiming the same route must never silently
        # overwrite or double-count by file order — the combined workbook is
        # not written.
        route = "" if route is None else str(route).strip()
        if not route:
            return ConsolidateResult(
                status="error",
                message=(f"{p.name} produced no usable route identity; the "
                         f"combined {report_name} workbook was not written. "
                         "Re-export that route's PDF, then run again."),
            )
        if route in route_sources:
            return ConsolidateResult(
                status="error",
                message=(f"Two PDFs both convert to route {route}: "
                         f"{route_sources[route]} and {p.name} (is the same "
                         f"route in the folder twice?). The combined "
                         f"{report_name} workbook was not written; remove "
                         "the duplicate and run again."),
            )
        route_sources[route] = p.name
        out_file = conv / f"{converted_prefix}_route_{route}.xlsx"
        candidate = conv / f".{converted_prefix}.tmp-{uuid.uuid4().hex}.xlsx"
        if not guarded(candidate) or not guarded(out_file):
            return destination_changed()
        try:
            # Serialize to an unpredictable sibling first, then revalidate the
            # exact conversion directory and final path at the commit boundary.
            write_one(rows, candidate)
            if not guarded(candidate) or not guarded(out_file):
                return destination_changed()
            os.replace(candidate, out_file)
        except PermissionError:
            if guarded(candidate):
                try:
                    candidate.unlink()
                except OSError:  # silent-ok: the user-facing locked-file error is primary
                    pass
            return ConsolidateResult(
                status="error",
                message=(f"Could not save {out_file.name}.\n\n"
                         "The file is probably open in Excel. Close it and try again."),
            )
        if not guarded(out_file):
            return destination_changed()
        events.on_log(f"  route {route}: {len(rows)} rows -> {out_file.name}")
        converted += 1

    if converted == 0:
        return ConsolidateResult(
            status="error",
            message=(f"None of the PDFs in:\n{in_dir}\n\ncontained readable "
                     f"{report_name} data. {unreadable_hint}"),
        )

    events.on_log("")

    # Combine all converted per-route files with the shared XLSX core. P12 TOCTOU:
    # pass the REAL confirm + the existence we saw at the early prompt down into
    # consolidate_xlsx so its pre-replace gate (atomic_save_if) catches a destination
    # that APPEARED after that prompt — at the final os.replace — without
    # re-prompting for the already-confirmed pre-existing case.
    # CMP-AUD-066: every combined workbook this driver writes is BY
    # CONSTRUCTION a PDF conversion, so the provenance marker rides the
    # report's own decoration.
    base_decorate = consolidate_kwargs.get("decorate_workbook")

    def _decorate_with_marker(wb):
        if base_decorate is not None:
            base_decorate(wb)
        write_pdf_source_marker(wb)

    consolidate_kwargs = dict(consolidate_kwargs,
                              decorate_workbook=_decorate_with_marker)
    result = consolidate_xlsx(
        input_dir=conv, out_path=out,
        events=events, confirm_overwrite=confirm, existed_at_confirm=existed_at_confirm,
        commit_guard=guarded,
        **consolidate_kwargs)
    if result.status == "ok":
        failed = ctx["failed"]
        result.summary_lines = [
            f"Route PDFs:   {len(pdfs) - len(failed)} converted"
            + (f", {len(failed)} failed {failed}" if failed else ""),
            f"Route files:  {converted} (in {conv})",
        ] + result.summary_lines
        # RR2-B1 / D18: every converted per-route file may have combined cleanly
        # (consolidate_xlsx -> complete), yet the PDF->row parse dropped data.
        # Those losses are NOT visible to the XLSX consolidator, so the report's
        # finalize ESCALATES to a producer-owned partial — a lossy output must
        # not be promoted / cached / compared as complete.
        finalize(result, ctx)
        _clamp_input_counts(result, ctx["pdfs"], events)
    return result


def _clamp_input_counts(result, discovered, events):
    """CMP-AUD-064 invariant: the file-level skipped/failed counts can never exceed
    the number of discovered input PDFs. A producer that leaks a line-level anomaly
    count into a file field is clamped and logged here, so no consumer of the
    structured outcome ever sees an impossible source count (e.g. three skips from
    one PDF with three malformed lines). Line-level parse anomalies belong in the
    ⚠ note + producer_extra['parse_anomalies'], never the file channels."""
    for field_name in ("skipped_inputs", "failed_inputs"):
        v = getattr(result, field_name)
        if isinstance(v, int) and v > discovered:
            events.on_log(f"  NOTE: {field_name}={v} exceeds {discovered} discovered "
                          f"PDF(s) — clamping to {discovered} (a line-level anomaly "
                          "count must not fill a file-count field).")
            setattr(result, field_name, discovered)
