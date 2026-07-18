"""Shared substrate for the five TSMIS-vs-TSN FILE comparators (P5b / S04).

`compare_ramp_detail_tsn`, `compare_ramp_summary_tsn`, `compare_highway_sequence_tsn`,
`compare_intersection_detail_tsn`, and `compare_intersection_summary_tsn` each wrote
the SAME `compare()` skeleton — deps gate, path/existence checks, the log banner, the
load try/except, then a `compare_core.run_compare` call — differing only in the schema,
the per-report loaders, the banner text, and (for the two FLAT detail reports) a couple
of normalizers + a Notes legend sheet. This module is that shared skeleton, so each
comparator is reduced to its **schema + projector**:

  * `run_files_compare` — the registry "files"-kind `compare()` driver. The report
    supplies a `loader(tsmis_path, tsn_path) -> (rows_t, rows_n, warnings)` (it may raise
    ValueError for a bad input shape) and the static facts (schema, banner, has_route,
    deps gate); the driver owns the boilerplate and the `run_compare` hand-off.
  * `make_notes_writer` — the identical "Notes" legend sheet builder the two FLAT detail
    comparators share (title + body lines differ; styling is fixed).
  * `norm_pm` / `iso_date` — the postmile + date normalizers Ramp Detail and Intersection
    Detail share verbatim (`iso_date` also handles Intersection Detail's 2-digit TSN year).

Behavior-neutral: the strings, branch order, and `run_compare` arguments match the
per-module bodies this replaced (the five golden `check_compare_*_tsn.py` canaries are
the semantic-identity proof). The comparison engine stays in `compare_core` — this
module never touches it. Console-free; openpyxl is imported lazily (only inside the
Notes writer, which runs solely when a workbook is actually being built).
"""
import contextlib
import hashlib
import json
import logging
import os
import re
from collections import Counter
from dataclasses import replace
from datetime import date
from pathlib import Path

import artifact_store
import consolidation_meta
import outcome
from compare_core import run_compare
from comparison_contract import comparison_result_boundary
from events import ConsolidateResult, Events
from paths import today_str

log = logging.getLogger(__name__)


# --------------------------------------------------------------------------- #
# shared row/name helpers (v0.19.0 R1 — the idioms every comparator copied)
# --------------------------------------------------------------------------- #
def row_has_data(row):
    """True when the row holds at least one non-blank cell — the emptiness
    predicate every loader wrote inline (`any(c is not None and str(c).strip()
    != "" ...)`); one spelling so they can't drift."""
    return bool(row) and any(c is not None and str(c).strip() != "" for c in row)


def require_exact_raw_header(header, expected, report_name):
    """Require one report edition's complete, ordered raw-header contract.

    Raw TSN workbooks are truth inputs, not best-effort tables: projecting by a
    sparse name dictionary would otherwise turn every absent compared source
    column into a clean-looking blank.  Equality here is deliberately exact --
    missing, duplicate, renamed, reordered, or extra columns all fail before a
    data row is projected.  The diagnostic calls out the first useful class of
    drift while remaining deterministic for hermetic checks.
    """
    got = list(header or [])
    want = list(expected)
    if got == want:
        return

    duplicates = []
    seen = set()
    for value in got:
        if value in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(value)
    missing = [value for value in want if value not in got]
    unexpected = [value for value in got if value not in want]
    if duplicates:
        reason = "duplicate column(s): " + ", ".join(repr(v) for v in duplicates)
    elif missing or unexpected:
        pieces = []
        if missing:
            pieces.append("missing " + ", ".join(repr(v) for v in missing))
        if unexpected:
            pieces.append("unexpected " + ", ".join(repr(v) for v in unexpected))
        reason = "; ".join(pieces)
    else:
        mismatch = next((i for i, pair in enumerate(zip(got, want))
                         if pair[0] != pair[1]), min(len(got), len(want)))
        reason = f"column order differs at position {mismatch + 1}"
    raise ValueError(
        f"the TSN {report_name} raw header does not match the complete "
        f"{len(want)}-column source schema ({reason})")


@contextlib.contextmanager
def exact_raw_rows(path, sheet_name, expected_header, report_name,
                   *, required_nonblank=()):
    """Yield rows from one exact, literal TSN statewide workbook edition.

    Raw truth is one physical document, not any sheet with familiar headings.
    Require the sole worksheet to be the named, visible source sheet; inspect
    formulas/errors rather than cached values; require the ordered header; and
    fail rows whose mandatory identity claims are blank. Optional identity
    components remain literal blanks and are never invented here.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        if wb.sheetnames != [sheet_name]:
            got = ", ".join(repr(name) for name in wb.sheetnames) or "no worksheets"
            raise ValueError(
                f"the TSN {report_name} raw workbook must contain exactly one "
                f"worksheet named {sheet_name!r} (found {got})")
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            raise ValueError(
                f"the TSN {report_name} raw worksheet {sheet_name!r} must be visible")

        cells = ws.iter_rows()
        header_cells = list(next(cells, ()) or ())
        for cell in header_cells:
            if cell.data_type in {"f", "e"}:
                raise ValueError(
                    f"the TSN {report_name} raw workbook contains a formula/error "
                    f"cell at {cell.coordinate}")
        header = [cell.value for cell in header_cells]
        require_exact_raw_header(header, expected_header, report_name)
        indices = {name: header.index(name) for name in required_nonblank}

        def rows():
            for row_number, raw_cells in enumerate(cells, start=2):
                raw_cells = list(raw_cells)
                values = [cell.value for cell in raw_cells]
                if not row_has_data(values):
                    continue
                for cell in raw_cells:
                    if cell.data_type in {"f", "e"}:
                        raise ValueError(
                            f"the TSN {report_name} raw workbook contains a "
                            f"formula/error cell at {cell.coordinate}")
                missing = [name for name, index in indices.items()
                           if index >= len(values)
                           or values[index] is None
                           or str(values[index]).strip() == ""]
                if missing:
                    raise ValueError(
                        f"the TSN {report_name} raw row {row_number} is missing "
                        f"required identity claim(s): {', '.join(missing)}")
                yield values

        yield header, rows()
    finally:
        wb.close()


_ROUTE_TOKEN_RE = re.compile(r"route[ _-]*([0-9]+[A-Za-z]?)", re.IGNORECASE)


def suggest_route_name(path, fallback_tag, name_tag):
    """Output-filename suggestion shared by the route-aware comparators:
    '<name_tag>_<RouteN|Consolidated|fallback_tag>_Comparison <today>.xlsx'.
    The trailing generated-on date stamps when the comparison was built (A1)."""
    stem = Path(path).stem
    m = _ROUTE_TOKEN_RE.search(stem)
    tag = (f"Route{m.group(1).lstrip('0') or '0'}" if m
           else "Consolidated" if "consolidated" in stem.lower() else fallback_tag)
    return f"{name_tag}_{tag}_Comparison {today_str()}.xlsx"


def require_per_route_identity(path_a, path_b):
    """CMP-AUD-049 (direct-compare half): a PER-ROUTE pair must demonstrably
    describe the SAME route. Per-route workbooks carry no Route column, so
    the only available identity is the filename's route token (every export
    and conversion this app writes carries one, e.g. "…route_005S.xlsx");
    both files must carry one and they must normalize equal. Two identically
    populated files named Route 001 and Route 002 used to be accepted as one
    universe and certified as a match. Consolidated pairs are content-keyed
    by their Route column and never come through here."""
    from pdf_table_lib import norm_route

    routes = []
    for path in (path_a, path_b):
        name = Path(path).name
        m = _ROUTE_TOKEN_RE.search(Path(path).stem)
        if not m:
            raise ValueError(
                f"{name} doesn't carry a route token in its filename (like "
                "'route 001'), so a per-route comparison can't verify both "
                "files describe the same route. Rename the file to include "
                "its route, or compare two consolidated workbooks.")
        routes.append(norm_route(m.group(1)))
    if routes[0] != routes[1]:
        raise ValueError(
            f"The two files name different routes: {Path(path_a).name} says "
            f"route {routes[0]} but {Path(path_b).name} says route "
            f"{routes[1]} — a per-route comparison must compare the same "
            "route. Pick matching per-route files, or two consolidated "
            "workbooks.")


def validate_route_universe(routes, name, workbook_path, note_sink=None):
    """CMP-AUD-183/071: an aggregated TSMIS summary's route universe must be
    sound before its category cells are summed — every data row carries a usable
    route identity and no route appears twice — and, when the producer recorded
    its ordered `route_census` in the outcome sidecar, must match that census
    EXACTLY (a dropped, added, renamed, reordered, or suffix-collapsed row
    refuses). A census-less workbook (older consolidation, or a copy without its
    sidecar) keeps the internal checks and gets an explicit no-census diagnostic
    instead of a silent pass. Shared by the Intersection Summary and Ramp Summary
    aggregate comparators (identical California route-id shape). `routes` is the
    [(row_no, cell_value)] the loader collected from the Route column."""
    cleaned = []
    for row_no, rv in routes:
        s = ("" if rv is None else str(rv)).strip()
        if isinstance(rv, bool) or not s or not s.isalnum() or len(s) > 8:
            raise ValueError(f"{name}: sheet row {row_no} has no usable route "
                             f"identity ({rv!r}) — its counts cannot be attributed")
        cleaned.append(s)
    if not cleaned:
        raise ValueError(f"{name} has no route rows — nothing to aggregate")
    dupes = sorted(r for r, n in Counter(cleaned).items() if n > 1)
    if dupes:
        raise ValueError(f"{name}: route(s) {', '.join(dupes[:6])} appear on more "
                         "than one row — refusing to aggregate an ambiguous "
                         "route universe")
    census = consolidation_meta.read_extra(workbook_path, "route_census")
    if census is None:
        note = (f"TSMIS route universe: {len(cleaned)} routes; no producer route "
                "census recorded (older consolidation) — internal checks only.")
    elif (not isinstance(census, list)
          or not all(isinstance(r, str) and r for r in census)):
        raise ValueError(f"{name}: the producer route census beside the workbook "
                         "is malformed — re-consolidate before comparing")
    elif census != cleaned:
        detail = (f"workbook has {len(cleaned)} route rows, the census records "
                  f"{len(census)}")
        for i, (want, got) in enumerate(zip(census, cleaned)):
            if want != got:
                detail = (f"first divergence at row {i + 2}: census {want!r} vs "
                          f"workbook {got!r}")
                break
        raise ValueError(f"{name}: the aggregated routes do not match the "
                         f"producer's route census ({detail}) — a route was "
                         "dropped, added, renamed, or reordered after "
                         "consolidation; re-consolidate before comparing")
    else:
        note = (f"TSMIS route universe verified against the producer census: "
                f"{len(cleaned)} routes ({cleaned[0]}–{cleaned[-1]}).")
    if note_sink is not None:
        note_sink.append(note)


def require_pdf_source(path, side_label, report_noun):
    """CMP-AUD-066 (PDF-role half): the "TSMIS (PDF)" side of a comparison
    must actually BE one of this app's PDF conversions — those workbooks carry
    a very-hidden versioned provenance marker (pdf_table_lib). An Excel
    consolidation has the same visible shape, so shape alone certified
    Excel-vs-Excel runs as PDF-vs-Excel. Unmarked (incl. pre-marker) picks
    refuse with a re-consolidate hint; a malformed marker refuses too (it
    cannot certify a version)."""
    from pdf_table_lib import pdf_source_marker_state

    if pdf_source_marker_state(path) < 1:
        raise ValueError(
            f"{Path(path).name} was not produced by this app's {report_noun} "
            f"(PDF) conversion (it has no valid PDF-conversion marker), so it "
            f"cannot stand as the {side_label} side — its rows may come from "
            f"the Excel export. Consolidate the {report_noun} (PDF) report "
            "(re-consolidate if this file predates the marker), then pick "
            "that workbook.")


def reject_pdf_source(path, side_label, report_noun):
    """CMP-AUD-066 (the mirror): the "TSMIS (Excel)" side must NOT be a
    PDF-sourced workbook — comparing a PDF conversion against itself would
    certify a 'PDF vs Excel' match no Excel export ever entered. Any marker
    presence (valid or malformed) refuses."""
    from pdf_table_lib import pdf_source_marker_state

    if pdf_source_marker_state(path) != 0:
        raise ValueError(
            f"{Path(path).name} is one of this app's {report_noun} (PDF) "
            f"conversions (it carries the PDF-conversion marker), so it "
            f"cannot stand as the {side_label} side. Pick the consolidated "
            f"{report_noun} workbook built from the Excel exports.")


# --------------------------------------------------------------------------- #
# In-workbook normalization marker (CMP-AUD-037) — the DIRECT-path freshness gate
# for the XLSX-sourced TSN families (Ramp Detail / Intersection Detail / Highway
# Detail). The matrix/library path already refuses a stale library via its
# certificate (report_catalog's normalization_version, D2), but a classic file
# comparison trusted ANY workbook carrying the normalized sheet — so a library
# built by an older normalizer was silently compared, resurrecting whatever that
# version got wrong. The consolidator families (HSL/HL) carry their own marker
# sheet; this is the shared mechanics the loader families reuse. The normalized
# rows sheet keeps its WIDTH across a marker-only bump, so the marker is the only
# reliable signal on a bare direct-path file.
# --------------------------------------------------------------------------- #
NORMALIZATION_MARKER_SHEET = "TSN Normalization"
_NORM_VERSION_LABEL = "Normalization version"


def write_normalization_marker(wb, version, *, report_name=None):
    """Stamp a normalized TSN workbook with its normalization `version`
    (CMP-AUD-037). Uses create_sheet + append so it works on the write-only
    normalized workbook too (`ws['A1'] =` TypeErrors in write-only mode). The
    marker is a plain trailing sheet; the data sheet is untouched."""
    ws = wb.create_sheet(NORMALIZATION_MARKER_SHEET)
    if report_name:
        ws.append(["Report", report_name])
    ws.append([_NORM_VERSION_LABEL, int(version)])


def normalization_marker_version(wb):
    """The declared normalization version from an OPEN workbook's marker sheet,
    or 0 when the sheet is absent OR malformed. Fail-safe by design: a 0 means
    the caller refuses with the rebuild hint, so a corrupt marker never passes
    as current."""
    if NORMALIZATION_MARKER_SHEET not in wb.sheetnames:
        return 0
    for r in wb[NORMALIZATION_MARKER_SHEET].iter_rows(values_only=True):
        if r and str(r[0]).strip() == _NORM_VERSION_LABEL:
            try:
                return int(r[1])
            except (TypeError, ValueError, IndexError):  # silent-ok: a malformed marker reads as version 0 — the caller then refuses with the rebuild hint (fail-safe)
                return 0
    return 0


def require_current_normalization(wb, name, version, detail):
    """Refuse an OPEN normalized TSN workbook older than `version` on the DIRECT
    comparison path (CMP-AUD-037). `detail` names what a pre-current file is
    missing; `name` is the display filename. A shape-stable marker-only bump
    means every width/label gate still passes a stale library — only the marker
    distinguishes it, so this is the authoritative freshness check."""
    if normalization_marker_version(wb) < version:
        raise ValueError(
            f"{name} was built by an older TSN converter ({detail}) — rebuild "
            "the TSN library and pick the fresh normalized workbook.")


def require_shared_header_prefix(header, expected_prefix, sidecars, name, report):
    """CMP-AUD-033: bind a normalized-library sheet to its EXACT shared-header
    prefix before any row is read. The normalized loaders discard the header and
    read cells BY POSITION, so a semantically REORDERED or RENAMED header would
    silently mis-map every column (a reordered workbook loaded PM as a county
    code). `expected_prefix` is the ordered ['Route'] + SHARED_HEADER the loader
    reads by position; `sidecars` is the exact set of documented columns that
    must follow it (order-independent, once each — the normalizer always writes
    all of them, and only them). Rejects a header whose prefix is missing,
    renamed, reordered, or duplicated, or whose trailing columns are not exactly
    the documented sidecars. Case/whitespace on the cell text is tolerated (the
    header is stripped); the column NAMES and ORDER must match."""
    hdr = [("" if c is None else str(c).strip()) for c in header]
    prefix = [str(c).strip() for c in expected_prefix]
    n = len(prefix)
    if hdr[:n] != prefix:
        raise ValueError(
            f"{name} is not a current normalized TSN {report} library — its "
            f"column layout does not match the expected order ('{prefix[0]}', "
            f"'{prefix[1]}', ...), so reading it by position would mis-map every "
            "field — rebuild the TSN library and pick the fresh normalized "
            "workbook.")
    trailing = [c for c in hdr[n:] if c]
    if sorted(trailing) != sorted(str(c).strip() for c in sidecars):
        raise ValueError(
            f"{name} does not carry exactly the documented TSN {report} sidecar "
            f"columns {sorted(str(c).strip() for c in sidecars)} (found "
            f"{sorted(trailing)}) — rebuild the TSN library and pick the fresh "
            "normalized workbook.")


def exact_consolidated_header_ok(*expected):
    """CMP-AUD-034: build the `header_ok` predicate that binds a CONSOLIDATED
    TSMIS workbook to its EXACT documented header. The consolidated `_load_tsmis`
    loaders read every field BY POSITION (the site's export labels are
    column-shifted against their values), so the previous width/last-label/`PM`-
    somewhere gates let a shifted, junk-relabelled, or wrong-report header be
    interpreted as the intended schema — yielding false differences, false
    one-sided rows, or a false match. Each `expected` is a full ordered
    ['Route', ...] signature `load_consolidated_rows` presents (each cell
    `str(c).strip()`, `""` for None; `header[0]=='Route'` is also checked there).
    An exact match rejects any insertion, deletion, block shift, or relabel.

    Pass MORE THAN ONE header when a report has more than one valid site EDITION
    with the SAME value POSITIONS but different column LABELS — the by-position
    loaders read either edition correctly, so both are accepted (backward
    compatibility) while the exact-label gate still rejects junk. Intersection
    Detail is the case: the 2026-07-17 site build corrected its long-misaligned
    header labels (`P`->`PP`, `S`->`PS`, the INT Type/Eff-Date pair realigned to
    their own values, `Xing P/S`->`Int PS`) WITHOUT moving any data, so the legacy
    (7.8/7.9) and current (2026-07-17) headers are both valid. A genuinely different
    layout (a real column insert/delete/shift, or a pre-July-2026 edition) still
    matches no accepted header and is refused. The header is data-source/env-
    independent, and each report's PDF-consolidated workbook carries a header the
    by-position loaders read the same way, so the polymorphic `_load_tsmis` (Highway
    Sequence/Detail + Intersection Detail load both shapes through one loader) stays
    valid across editions."""
    exps = [[("" if c is None else str(c).strip()) for c in e] for e in expected]
    return lambda header: [("" if c is None else str(c).strip())
                           for c in header] in exps


def source_files_from_consolidated(path, sheet, prefix, ext="xlsx"):
    """The per-row source export filename `<prefix>_route_<route>.<ext>`, one per
    data row, read from the CONSOLIDATED workbook's leading `Route` column (col 0)
    in the SAME order and with the SAME `row_has_data` filter the loaders use — so
    it index-aligns with the transformed rows. The FILE route (col 0) is used, not a
    downstream re-derivation, so for Intersection Detail it names the export a row
    actually came from even when the compared route differs (a junction/equate row).
    Returns [] if the workbook/sheet can't be read."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log.warning("source-file provenance: could not open %s (%s: %s); the "
                    "Source Files sheet is omitted", Path(path).name,
                    type(e).__name__, str(e).splitlines()[0] if str(e) else "")
        return []
    try:
        if sheet not in wb.sheetnames:
            return []
        it = wb[sheet].iter_rows(values_only=True)
        header = list(next(it, []))
        # Only a CONSOLIDATED workbook (a prepended leading 'Route' column) carries
        # the per-route file identity in column 0. A bare per-route export doesn't,
        # so we skip it rather than name a data cell as a route.
        if not header or str(header[0]).strip() != "Route":
            return []
        out = []
        for r in it:
            if row_has_data(r):
                route = "" if not r or r[0] is None else str(r[0]).strip()
                out.append(f"{prefix}_route_{route}.{ext}" if route else "")
        return out
    finally:
        wb.close()


def source_files_from_rows(rows, prefix, ext):
    """Per-row source filename `<prefix>_route_<route>.<ext>` from the route already
    prepended at column 0. For sides loaded straight from per-route files with no
    consolidated workbook to read (the cross-environment / baseline path, where both
    sides are per-route TSMIS exports and the loader prepends the route)."""
    out = []
    for r in rows:
        route = "" if not r or r[0] is None else str(r[0]).strip()
        out.append(f"{prefix}_route_{route}.{ext}" if route else "")
    return out


def write_source_files_sheet(wb, side_specs, sheet_title="Source Files"):
    """Append a companion "Source Files" sheet documenting which per-route export
    each TSMIS row came from. `side_specs` = [(side_name, rows, files), ...] where
    `rows` are that side's transformed rows and `files` the index-aligned source
    filenames (from `source_files_from_consolidated`). Write-only-safe (create_sheet
    + append plain values), so it works with the streaming comparison workbook and
    needs no change to the correctness-locked engine. Rows the side actually wrote
    are listed in order, so row N here corresponds to data-sheet row N+1."""
    ws = wb.create_sheet(sheet_title)
    ws.append(["Side", "Row #", "Route (as compared)", "Source File"])
    for side_name, rows, files in side_specs:
        for i, (row, fn) in enumerate(zip(rows, files), start=1):
            route = "" if not row or row[0] is None else str(row[0])
            ws.append([side_name, i, route, fn])


def with_source_files_sheet(schema, path_a, path_b):
    """Return `schema` (a copy) whose `extra_sheet_writer` ALSO writes the default
    "Source Files" companion sheet, driven by the schema's `source_file_a`/`_b`
    (prefix, sheet, ext) specs and the two input PATHS. Composes with any existing
    extra_sheet_writer (Report View, the Ramp Summary rollup, …). This is the single
    place the file-based comparison substrate wires provenance, so every report gets
    the sheet by default without its own writer. No-op (schema unchanged) when
    neither side declares a spec (e.g. a report that opts out, or the statewide-TSN
    side)."""
    spec_a = getattr(schema, "source_file_a", ())
    spec_b = getattr(schema, "source_file_b", ())
    if not spec_a and not spec_b:
        return schema
    prev = schema.extra_sheet_writer

    def _writer(wb, ctx):
        if prev is not None:
            prev(wb, ctx)
        specs = []
        for spec, path, side, rows_key in (
                (spec_a, path_a, schema.side_a, "rows_a"),
                (spec_b, path_b, schema.side_b, "rows_b")):
            if spec:
                prefix, sheet, ext = spec
                files = source_files_from_consolidated(str(path), sheet, prefix, ext)
                if files:
                    specs.append((side, ctx[rows_key], files))
        if specs:
            write_source_files_sheet(wb, specs)

    return replace(schema, extra_sheet_writer=_writer)


def load_consolidated_rows(path, sheet_name, *, missing_sheet_hint, bad_header_msg,
                           header_ok=None, row_transform=list):
    """The consolidated-workbook loader skeleton three vs-TSN comparators wrote
    verbatim: open (user-safe ValueError on failure) -> require `sheet_name` ->
    read + strip the header -> demand a leading 'Route' column (plus the
    report's `header_ok(header)` drift guard) -> the non-empty data rows through
    `row_transform(list(row))`. Returns `(rows, True)` — the consolidated shape
    is always route-keyed. openpyxl is imported here (lazily), matching the
    module's deps posture."""
    from openpyxl import load_workbook

    name = Path(path).name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {name}: {type(e).__name__}: {e}")
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{name} has no '{sheet_name}' sheet — {missing_sheet_hint}")
        it = wb[sheet_name].iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in (next(it, []) or [])]
        if (not header or header[0] != "Route"
                or (header_ok is not None and not header_ok(header))):
            raise ValueError(f"{name} {bad_header_msg}")
        return [row_transform(list(r)) for r in it if row_has_data(r)], True
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# TSN print identity (CMP-AUD-146) — the statewide summary prints' report
# identity/timing/submitter facts, extracted from the FULL document text.
# --------------------------------------------------------------------------- #
# field -> regex over the joined page texts. A field may print on several pages
# (report id / event id repeat on the data-page header) but must carry exactly
# ONE distinct value; page-1 policy prose is legal furniture, not identity.
_IDENTITY_PATTERNS = {
    "report_id": r"\b(OTM\d+)\b",
    "report_date": r"REPORT DATE\s*:\s*(\d{2}/\d{2}/\d{4})",
    "reference_date": r"REFERENCE DATE\s*:\s*(\d{2}/\d{2}/\d{4})",
    "submitter": r"SUBMITTOR\s*:\s*(\S+)",
    "report_title": r"REPORT TITLE\s*:\s*'\s*(.*?)\s*'",
    # The value can sit on the NEXT line ("EVENT ID :" / "4843742" on the Ramp
    # print) and prints colon-less on the data page ("Event ID 4843738").
    "event_id": r"EVENT ID\s*:?\s*\n?\s*(\d+)\b",
    "generated_time": r"\b(\d{2}:\d{2} [AP]M)\b",
    "location_criteria": r"LOCATION CRITERIA:\s*\n\s*([^\n]+)",
}


def tsn_print_identity(full_text, source):
    """{field: value} for the statewide TSN print's report identity (CMP-AUD-146):
    report id, report/reference dates, submitter, title, event id, generation
    time, location criteria. Every field is REQUIRED and must resolve to exactly
    one distinct value — a print we cannot identify (or that carries conflicting
    identities) refuses instead of normalizing anonymous data."""
    out = {}
    problems = []
    for field, pattern in _IDENTITY_PATTERNS.items():
        values = {m.strip() for m in re.findall(pattern, full_text,
                                                re.IGNORECASE | re.MULTILINE)}
        if not values:
            problems.append(f"{field} not found")
        elif len(values) > 1:
            problems.append(f"{field} has conflicting values: "
                            + ", ".join(sorted(values)[:4]))
        else:
            out[field] = next(iter(values))
    if problems:
        raise ValueError(f"{source}: the TSN print's identity could not be "
                         "established — " + "; ".join(problems))
    return out


# --------------------------------------------------------------------------- #
# Durable comparison provenance (CMP-AUD-076) — the saved artifact must be able
# to say exactly WHAT it compared: the full canonical selections (basenames are
# ambiguous — A\same.xlsx vs B\same.xlsx), each input's pre-read content digest
# and stat identity, the producer's recorded completion, and the recipe facts.
# Persisted as a tolerant `.provenance.json` sidecar beside the workbook (the
# write_outcome pattern: guard-disciplined temp+replace; absence reads as an
# older comparison, never a silent pass). Folding this into the strict schema-v4
# payload is Phase-5 artifact-epoch work.
# --------------------------------------------------------------------------- #
PROVENANCE_SUFFIX = ".provenance.json"
_PROVENANCE_SCHEMA = 1
_SHA_CHUNK = 1024 * 1024


def _sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(_SHA_CHUNK), b""):
            h.update(chunk)
    return h.hexdigest()


def capture_input_provenance(sides):
    """[{role, name, selection, sha256, size, mtime_ns}] for ``sides`` =
    ((role_label, path), ...), hashed BEFORE any loader reads (CMP-AUD-098
    discipline: the record binds what the comparison actually consumed; the
    publication boundary's current-source recheck refuses a mid-run swap).
    Raises ValueError when an input can't be read — a comparison whose inputs
    can't be identified must not run."""
    captured = []
    for role, raw in sides:
        path = Path(raw)
        try:
            st = path.stat()
            digest = _sha256_file(path)
        except OSError as e:
            raise ValueError(f"Could not capture the {role} input identity for "
                             f"{path.name}: {type(e).__name__}: {e}")
        producer = consolidation_meta.read_outcome(path)
        captured.append({
            "role": role,
            "kind": "file",
            "name": path.name,
            "selection": str(path.resolve(strict=False)),
            "sha256": digest,
            "size": st.st_size,
            "mtime_ns": st.st_mtime_ns,
            "producer_completion": (producer.completion if producer is not None
                                    else None),
        })
    return captured


def provenance_path(workbook_path):
    p = Path(workbook_path)
    return p.with_name(p.name + PROVENANCE_SUFFIX)


def write_comparison_provenance(result, out_path, *, report, banner, inputs,
                                commit_guard=None):
    """Persist the provenance sidecar beside a JUST-COMMITTED comparison.
    Additive evidence, never a gate: only a result that committed real bytes
    (an attached artifact generation) gets a record; a write failure logs and
    leaves the comparison result intact (absent provenance reads as an older
    comparison). Guard discipline mirrors consolidation_meta.write_outcome."""
    generation = getattr(result, "artifact_generation", None)
    if getattr(result, "status", None) != "ok" or generation is None:
        return False
    target = provenance_path(out_path)
    tmp = target.with_name(target.name + ".tmp")
    if not (consolidation_meta.guard_allows(commit_guard, Path(out_path))
            and consolidation_meta.guard_allows(commit_guard, target)
            and consolidation_meta.guard_allows(commit_guard, tmp)):
        log.warning("comparison provenance for %s: destination changed; "
                    "no sidecar write", Path(out_path).name)
        return False
    payload = {
        "schema_version": _PROVENANCE_SCHEMA,
        "recipe": {"report": report, "banner": banner},
        "inputs": inputs,
        "generation_id": generation.generation_id,
        "members": dict(generation.content_digests),
    }
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f)
        if not (consolidation_meta.guard_allows(commit_guard, Path(out_path))
                and consolidation_meta.guard_allows(commit_guard, target)):
            return False
        os.replace(tmp, target)
        return True
    except OSError as e:
        log.warning("comparison provenance for %s could not be written "
                    "(%s: %s)", Path(out_path).name, type(e).__name__, e)
        with contextlib.suppress(OSError):
            tmp.unlink()
        return False


def read_comparison_provenance(workbook_path):
    """The provenance record beside a comparison workbook, or None (absent /
    unreadable / wrong shape — an older comparison; never a fabricated one)."""
    try:
        with open(provenance_path(workbook_path), encoding="utf-8") as f:
            data = json.load(f)
    except OSError:      # silent-ok: absent sidecar is the pre-076 state
        return None
    except ValueError as e:
        log.info("comparison provenance beside %s unreadable (%s: %s)",
                 Path(workbook_path).name, type(e).__name__, e)
        return None
    return data if isinstance(data, dict) else None


# --------------------------------------------------------------------------- #
# shared normalizers (Ramp Detail + Intersection Detail, verbatim)
# --------------------------------------------------------------------------- #
def norm_pm(pm):
    """Postmile to one canon: strip the zero-padding TSN prints (' 000.606')
    while keeping the decimal ('0.606' stays distinct from '000.606'). A real
    numeric 0/0.0 canonicalizes to '0', NOT blank — `pm or ""` was the falsy-zero
    idiom (v0.18.3's phantom-diff root cause), and THIS normalizer feeds the
    row-ALIGNMENT keys, so a blanked 0 mis-aligned rows, not just cells."""
    s = ("" if pm is None else str(pm)).strip()
    if not s:
        return ""
    neg = s.startswith("-")
    s = s.lstrip("-").lstrip("0") or "0"
    if s.startswith("."):
        s = "0" + s
    return ("-" + s) if neg else s


def decimal_pm(pm):
    """The Decimal-canonical numeric postmile as text (CMP-AUD-006): leading
    zeros stripped (norm_pm), then trailing fraction zeros and a bare dot
    stripped — '9.6', '9.600', and '009.600' all identify the SAME point
    ('9.6'); '005.870' -> '5.87'; '001.000' -> '1'; '0', '0.0', '000.000' ->
    '0'. This is the IDENTITY component (what PhysicalIdentity equality
    hashes); the norm_pm text stays the display payload, so the two renders'
    printed forms remain visible while physically identical rows align."""
    s = norm_pm(pm)
    neg, t = s.startswith("-"), s.lstrip("-")
    if "." in t:
        ip, fp = t.split(".", 1)
        fp = fp.rstrip("0")
        t = ip + ("." + fp if fp else "")
    return ("-" + t) if neg and t not in ("", "0") else t


def _valid_ymd(yyyy, mm, dd):
    """True iff (yyyy, mm, dd) is a real calendar date (leap-day aware)."""
    try:
        date(yyyy, mm, dd)
        return True
    except ValueError:                    # silent-ok: date() IS the calendar-validity probe; invalid -> not a real date
        return False


def iso_date(d):
    """A Date of Record to YYYY-MM-DD across the formats the two systems print:
    TSMIS 'MM/DD/YYYY', TSN 'YYYY-MM-DD[ HH:MM:SS]', and TSN's 2-digit 'YY-MM-DD'
    (windowed at 30: >=30 -> 19xx, else 20xx).

    CMP-AUD-038: match the documented forms in FULL and construct the result only
    when the calendar date is real. Trailing corruption ('02/25/1976 junk') and
    impossible dates ('02/31/1976', non-leap '02/29/1900') are PRESERVED verbatim
    so they surface as a visible difference — never silently truncated to a clean
    date nor faked into a plausible-looking ISO string."""
    s = ("" if d is None else str(d)).strip()
    if not s:
        return ""
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)      # TSMIS 'MM/DD/YYYY'
    if m:
        mm, dd, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if _valid_ymd(yyyy, mm, dd):
            return f"{yyyy:04d}-{mm:02d}-{dd:02d}"
        return s
    m = re.fullmatch(                                        # TSN 'YYYY-MM-DD[ time]'
        r"(\d{4})-(\d{2})-(\d{2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?(?:\.\d+)?)?", s)
    if m:
        if _valid_ymd(int(m.group(1)), int(m.group(2)), int(m.group(3))):
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        return s
    m = re.fullmatch(r"(\d{2})-(\d{2})-(\d{2})", s)          # TSN '73-10-19' (YY-MM-DD)
    if m:
        yy = int(m.group(1))
        cc = 1900 if yy >= 30 else 2000                     # 2-digit-year window
        if _valid_ymd(cc + yy, int(m.group(2)), int(m.group(3))):
            return f"{cc + yy}-{m.group(2)}-{m.group(3)}"
        return s
    return s


# --------------------------------------------------------------------------- #
# producer completeness carried by selected workbook sidecars (Phase 2B)
# --------------------------------------------------------------------------- #
def _merge_input_outcomes(inputs, loader_warnings):
    """Merge current producer outcomes with loader diagnostics.

    ``inputs`` is ``((side_label, workbook_path), ...)``.  Only the coupled,
    mtime-validated :func:`consolidation_meta.read_outcome` record is consumed;
    summary prose and independently-read additive sidecar fields are never state.
    ``None`` (truly absent or demonstrably stale) preserves legacy/raw behavior
    without making a positive trust claim.

    Returns ``(warnings, input_completion, skipped, failed, failures)`` for the
    additive ``run_compare`` input contract.  A current failed/no-data/cancelled
    producer artifact is unusable and raises ``ValueError`` before comparison.
    """
    original_warnings = (() if loader_warnings is None else loader_warnings)
    merged_warnings = list(original_warnings)
    added_warning = False
    sidecar_skipped = 0
    sidecar_failed = 0
    structured_failures = []
    records = []

    for side, path in inputs:
        record = consolidation_meta.read_outcome(path)
        # The reader deliberately collapses absent and valid-but-stale to None.
        # A defensive current check also prevents a foreign stub from gaining trust.
        if record is None or not record.current:
            records.append(None)
            continue
        records.append(record)
        label = f"{side} input '{Path(path).name}'"

        if record.completion in (outcome.FAILED, outcome.NO_DATA, outcome.CANCELLED):
            raise ValueError(
                f"{label} has a current producer outcome of "
                f"'{record.completion}' and is not usable for comparison. "
                "Rebuild it or select a complete/partial workbook.")

        skipped = record.skipped_inputs
        failed = record.failed_inputs
        if (record.completion == outcome.COMPLETE and record.trusted
                and skipped == 0 and failed == 0):
            continue

        # Anything current that is not the exact trusted-complete shape is
        # incomplete. Valid partial records contribute exact counters; untrusted
        # records intentionally expose None rather than invented zeroes.
        if isinstance(skipped, int) and not isinstance(skipped, bool):
            sidecar_skipped += skipped
        if isinstance(failed, int) and not isinstance(failed, bool):
            sidecar_failed += failed

        if record.trusted:
            detail = (f"producer outcome is '{record.completion}' "
                      f"(skipped inputs: {skipped}; failed inputs: {failed})")
        else:
            detail = ("outcome metadata is current but untrusted"
                      + (f" ({record.diagnostic})" if record.diagnostic else ""))
        note = f"{label}: {detail}; comparison coverage is incomplete."
        merged_warnings.append(note)
        added_warning = True
        if (not record.trusted or (isinstance(failed, int) and failed > 0)):
            structured_failures.append(note)

    has_incomplete_record = added_warning
    all_current_complete = (len(records) == len(inputs) and bool(records)
                            and all(record is not None
                                    and record.trusted
                                    and record.completion == outcome.COMPLETE
                                    and record.skipped_inputs == 0
                                    and record.failed_inputs == 0
                                    for record in records))
    input_completion = (outcome.PARTIAL if has_incomplete_record
                        else outcome.COMPLETE if all_current_complete
                        else None)

    # Preserve the exact legacy warnings object/type when no sidecar diagnostic
    # was added. None then lets run_compare retain its historical len(warnings)
    # skipped count. Once a sidecar participates, pass exact combined counters so
    # the explanatory warning itself is not miscounted as another skipped input.
    warnings = merged_warnings if added_warning else original_warnings
    skipped_inputs = ((len(original_warnings) + sidecar_skipped)
                      if added_warning else None)
    return (warnings, input_completion, skipped_inputs, sidecar_failed,
            tuple(structured_failures))


# --------------------------------------------------------------------------- #
# same-source (PDF vs Excel) render-artifact equivalence — CMP-AUD-197 class
# --------------------------------------------------------------------------- #
# Owner ruling (2026-07-16, on the Intersection Detail PDF-vs-Excel report of
# eight "HILLCREST RD ≠ HILLCREST RD"-style cells): in a SAME-SOURCE self-check
# both sides render the SAME report, so a difference must be a data
# disagreement — an export-encoding artifact the other render structurally
# cannot carry is a false positive. Two censused artifact classes:
#   * the Excel export's OOXML control escapes ("…_x000d_" = an encoded CR per
#     installed Excel; the print renders nothing/whitespace there), and
#   * the Excel export's edge tab padding ("HILLCREST RD\t\t"; Excel TRIM
#     collapses SPACES only, so tabs survived to flag).
# Applied ONLY by the PDF-vs-Excel flavors, at their load boundary — every
# vs-TSN leg keeps its accepted oracle's byte-exact semantics (both machine
# formats genuinely carry these bytes there, and they compare equal).
_OOXML_ESCAPE_RE = re.compile(r"_x([0-9A-Fa-f]{4})_")
_EDGE_WHITESPACE = " \t\r\n\f\v"


def decode_ooxml_escapes(text):
    """Decode OOXML `_xHHHH_` escapes in a str (hex digits in either case;
    `_x005F_xHHHH_` decodes to the literal `_xHHHH_` per the OOXML spec,
    because the leftmost scan consumes the `_x005F_` first). Byte-equivalent
    to openpyxl.utils.escape.unescape — the Stage-8 oracles' xlsx reading —
    so vs-TSN loaders may share it (CMP-AUD-197's HSL half) without drifting
    from the accepted oracle semantics."""
    return _OOXML_ESCAPE_RE.sub(
        lambda m: chr(int(m.group(1), 16)), text)


def same_source_render_text(value):
    """One cell under same-source render equivalence: decode OOXML `_xHHHH_`
    escapes, map the decoded/whitespace-class characters' EDGES away, and keep
    interior breaks as separation (they collapse to one space, exactly how the
    print renders a wrapped value; compare_core's TRIM twin collapses space
    runs at compare time). Non-strings — including PhysicalKey cells, a str
    SUBCLASS the engine's identity rides on — pass through untouched."""
    if type(value) is not str:
        return value
    cleaned = re.sub(r"[\t\r\n\f\v]", " ", decode_ooxml_escapes(value))
    return cleaned.strip(_EDGE_WHITESPACE)


def same_source_render_rows(rows):
    """`same_source_render_text` over every cell of every loaded row (typed
    key cells pass through by construction)."""
    return [[same_source_render_text(cell) for cell in row] for row in rows]


# --------------------------------------------------------------------------- #
# shared Notes legend sheet (Highway Sequence + Intersection Detail)
# --------------------------------------------------------------------------- #
def make_notes_writer(title, lines, *, tab_color="ED7D31", col_width=110):
    """Return a `legend_writer(wb)` that appends the standard orange-tabbed "Notes"
    sheet: a filled white title row then one wrapped body row per line. Styling is
    fixed (Arial title/body, the 1F3864 fill, an A-column width); only `title` and
    `lines` vary per report. openpyxl is imported here (not at module import) so this
    module loads even where the comparators' deps are absent — the writer only ever
    runs while a workbook is being built (deps present)."""
    def _write(wb):
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Notes")
        ws.sheet_properties.tabColor = tab_color
        write_only = getattr(wb, "write_only", False)
        title_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        fill = PatternFill("solid", start_color="1F3864")
        body = Font(name="Arial", size=10)
        wrap = Alignment(vertical="top", wrap_text=True)

        def cell(value, font=body, f=None, align=None):
            if not write_only:
                return value
            c = WriteOnlyCell(ws, value=value)
            c.font = font
            if f:
                c.fill = f
            if align:
                c.alignment = align
            return c

        ws.column_dimensions["A"].width = col_width
        ws.append([cell(title, title_font, fill)])
        for line in lines:
            ws.append([cell(line, body, align=wrap)])
        return ws

    return _write


# --------------------------------------------------------------------------- #
# the shared compare() driver
# --------------------------------------------------------------------------- #
@comparison_result_boundary
def run_files_compare(schema, tsmis_path, tsn_path, out_path, *, banner, has_route,
                      loader, deps_ok=True,
                      deps_msg="Required components are missing (openpyxl).",
                      side_a="TSMIS", side_b="TSN",
                      events=None, confirm_overwrite=None, mode="formulas",
                      commit_guard=None):
    """The registry "files"-kind `compare()` skeleton shared by every file
    comparator: a deps gate -> path coercion + existence checks -> the log banner ->
    `loader(path_a, path_b)` (may raise ValueError for a bad input shape) ->
    `compare_core.run_compare`. `loader` returns `(rows_a, rows_b, warnings)`; a
    `warnings` of None is the `run_compare` default (no unreadable inputs). Two
    opt-in extensions (defaults = the original vs-TSN behavior, so the five P5b
    comparators are untouched):

      * `side_a`/`side_b` — the two side labels used in the existence-check
        message and the banner's file lines (the PDF-sourced flavors label their
        pickers "TSMIS (PDF)" / "TSMIS (Excel)").
      * `has_route=None` — the route-ness is DYNAMIC (per-route vs consolidated
        inputs); the loader then returns `(rows_a, rows_b, warnings, has_route)`.

    Returns a `ConsolidateResult`, the same contract the GUI/console drive
    identically."""
    events = events or Events()
    if not deps_ok:
        return ConsolidateResult(status="error", message=deps_msg)
    tsmis_path, tsn_path = Path(tsmis_path), Path(tsn_path)
    for p, side in ((tsmis_path, side_a), (tsn_path, side_b)):
        if not p.is_file():
            return ConsolidateResult(status="error",
                                     message=f"The {side} file doesn't exist:\n{p}")

    destinations = artifact_store.comparison_output_paths(out_path, mode)
    try:
        captured_sources = artifact_store.capture_source_identities(
            (tsmis_path, tsn_path))
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))

    def alias_error():
        try:
            artifact_store.ensure_outputs_do_not_alias_sources(
                destinations, (tsmis_path, tsn_path),
                captured_sources=captured_sources,
                require_sources_current=True)
        except ValueError as e:
            return ConsolidateResult(status="error", message=str(e))
        return None

    blocked = alias_error()
    if blocked is not None:
        return blocked

    # CMP-AUD-076: capture each input's full canonical selection + content
    # digest BEFORE any loader reads it, and log the full selections (the
    # banner keeps concise names; basenames alone are ambiguous).
    try:
        input_provenance = capture_input_provenance(
            ((side_a, tsmis_path), (side_b, tsn_path)))
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))

    events.on_log("=" * 60)
    events.on_log(banner)
    events.on_log("=" * 60)
    pad = max(len(side_a), len(side_b)) + 1
    events.on_log(f"{side_a + ':':<{pad}} {tsmis_path.name}")
    events.on_log(f"{side_b + ':':<{pad}} {tsn_path.name}")
    for rec in input_provenance:
        events.on_log(f"  {rec['role']} selection: {rec['selection']} "
                      f"(sha256 {rec['sha256'][:12]}…)")
    events.on_log("")

    try:
        loaded = loader(tsmis_path, tsn_path)
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))
    if has_route is None:
        rows_t, rows_n, warnings, has_route = loaded
    else:
        rows_t, rows_n, warnings = loaded

    # A selected consolidated workbook can be structurally readable while its
    # producer explicitly reported incomplete coverage. Preserve that truth in
    # every output flavor; absent/stale metadata remains the legacy/raw path.
    try:
        (warnings, input_completion, skipped_inputs, failed_inputs,
         input_failures) = _merge_input_outcomes(
            ((side_a, tsmis_path), (side_b, tsn_path)), warnings)
    except ValueError as e:
        return ConsolidateResult(status="error", message=str(e))

    # Loading can be long enough for a destination to appear or be redirected.
    # Recheck at the public write boundary so direct callers receive the same
    # protection as GUI/Matrix callers wrapped by commit_workbook.
    blocked = alias_error()
    if blocked is not None:
        return blocked

    prov_display = {"recipe": {"report": getattr(schema, "report_name",
                                                 str(schema)),
                               "banner": banner},
                    "inputs": input_provenance}
    # Default provenance: compose the "Source Files" companion sheet from the
    # schema's per-side source-file specs (no-op when neither side declares one).
    schema = with_source_files_sheet(schema, tsmis_path, tsn_path)
    committed = artifact_store.commit_workbook(
        out_path,
        lambda tmp: run_compare(
            schema, rows_t, rows_n, has_route, tmp,
            events=events, confirm_overwrite=lambda _p: True,
            mode=mode, name_a=tsmis_path.name, name_b=tsn_path.name,
            warnings=() if warnings is None else warnings,
            input_completion=input_completion,
            skipped_inputs=skipped_inputs,
            failed_inputs=failed_inputs,
            failures=input_failures,
            provenance=prov_display,
            commit_guard=commit_guard),
        twin=(mode == "both"), expect_sheet="Comparison",
        confirm_overwrite=confirm_overwrite,
        source_paths=(tsmis_path, tsn_path),
        captured_sources=captured_sources,
        commit_guard=commit_guard,
        requested_mode=mode)
    # CMP-AUD-076: bind the pre-read input identities to the committed
    # generation, beside the workbook (additive evidence, never a gate).
    write_comparison_provenance(committed, out_path,
                                report=getattr(schema, "report_name", str(schema)),
                                banner=banner, inputs=input_provenance,
                                commit_guard=commit_guard)
    return committed
