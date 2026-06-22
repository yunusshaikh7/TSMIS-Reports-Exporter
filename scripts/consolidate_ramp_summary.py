"""Consolidate TSAR: Ramp Summary PDFs into a single Excel workbook.

Reads every PDF in   output/ramp_summary/
Writes one workbook in output/consolidated/tsar_ramp_summary_consolidated.xlsx
with one row per route plus audit columns that verify the parsed numbers
sum to the reported total.

This script is self-contained on purpose: the parsing logic is specific
to the TSAR Ramp Summary PDF layout (two columns, wrapped descriptions,
orphan numbers) and is not reused by the other consolidators. Each
report's consolidator gets its own file under scripts/ so a layout
change in one report cannot break another.
"""
import logging
import re
from pathlib import Path

# pdfplumber wraps pdfminer.six, which logs a "Could not get FontBBox from
# font descriptor" warning for every font with a malformed bbox — these
# PDFs hit it on nearly every page. Parsing is unaffected; just silence it.
logging.getLogger("pdfminer").setLevel(logging.ERROR)

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from paths import (OUTPUT_ROOT, latest_output_day, output_day_dir,
                   stamped_consolidated_filename)
import outcome
import artifact_store
from events import Events, ConsolidateResult
from compare_core import is_formula_injection   # shared formula-injection guard

SUBDIR = "ramp_summary"
FILENAME = "tsar_ramp_summary_consolidated.xlsx"

# Legacy flat-layout locations (pre-dated exports); still used when no dated
# output/<YYYY-MM-DD>/ folders exist, so old exports stay consolidatable.
INPUT_DIR = OUTPUT_ROOT / SUBDIR
OUT_DIR = OUTPUT_ROOT / "consolidated"
OUT_PATH = OUT_DIR / FILENAME

# Friendly report name for user-facing messages (shown in both the GUI and
# the console, so keep these UI-neutral -- no ".bat" / "menu option" wording).
REPORT_NAME = "Ramp Summary"


def input_dir_for(day):
    """Per-route exports for `day` (a run-folder name); None = the legacy flat layout."""
    return (output_day_dir(day) / SUBDIR) if day else INPUT_DIR


def out_path_for(day):
    """Consolidated workbook destination for `day` (a run-folder name); None = the
    legacy location. The filename carries the run's date + source/environment (A1)
    so a copy lifted out of its folder keeps its provenance."""
    if not day:
        return OUT_PATH
    return output_day_dir(day) / "consolidated" / stamped_consolidated_filename(FILENAME, day)


# =============================================================================
# Schema — (column_name, label_match_regex), in report order
# =============================================================================

HIGHWAY_GROUPS = [
    ("hwy_right",         r"^R\s*-\s*Right$"),
    ("hwy_divided",       r"^D\s*-\s*Divided$"),
    ("hwy_undivided",     r"^U\s*-\s*Undivided$"),
    ("hwy_unconstructed", r"^X\s*-\s*Unconstructed$"),
    ("hwy_left",          r"^L\s*-\s*Left$"),
    ("hwy_others",        r"^Others$"),
]
ONOFF = [
    ("onoff_on",    r"^ON\s*-\s*On$"),
    ("onoff_off",   r"^OFF\s*-\s*Off$"),
    ("onoff_other", r"^OTH\s*-\s*Other$"),
]
POP_GROUPS = [
    ("pop_rural_inside",  r"^R-RURAL\s*-I\s*INSIDE CITY$"),
    ("pop_rural_outside", r"^-O\s*OUTSIDE CITY$"),     # 1st occurrence
    ("pop_urban_inside",  r"^U-URBAN\s*-I\s*INSIDE CITY$"),
    ("pop_urban_outside", r"^-O\s*OUTSIDE CITY$"),     # 2nd occurrence
    ("pop_invalid",       r"^-INVALID DATA$"),
]
RAMP_TYPES = [
    ("ramp_A_frontage",     r"^A\s*-\s*Frontage Road$"),
    ("ramp_B_collector",    r"^B\s*-\s*Collector Road$"),
    ("ramp_C_connector_L",  r"^C\s*-\s*Direct or Semi-direct Connector \(Left\)$"),
    ("ramp_D_diamond",      r"^D\s*-\s*Diamond Type Ramp$"),
    ("ramp_E_slip",         r"^E\s*-\s*Slip Ramp$"),
    ("ramp_F_connector_R",  r"^F\s*-\s*Direct or Semi-direct Connector \(Right\)$"),
    ("ramp_G_loop_left",    r"^G\s*-\s*Loop \(w/Left turn\)$"),
    ("ramp_H_buttonhook",   r"^H\s*-\s*Buttonhook Ramp$"),
    ("ramp_J_scissors",     r"^J\s*-\s*Scissors$"),
    ("ramp_K_split",        r"^K\s*-\s*Split Ramp$"),
    ("ramp_L_loop_noleft",  r"^L\s*-\s*Loop without Left Turn$"),
    ("ramp_M_two_way",      r"^M\s*-\s*Two way Ramp Segment$"),
    # P/V are statewide TSN ramp classifications. No per-route TSMIS PDF in the
    # 6.19 ground-truth set emits them, but TSN's statewide Ramp Summary does
    # (P=122, V=81), so the schema carries them in TSN document order (P after M,
    # V after R) to (a) match TSN exactly for the vs-TSN comparison and (b) capture
    # them should a TSMIS route ever report one, instead of silently dropping it.
    ("ramp_P_dummy_paired", r"^P\s*-\s*Dummy Paired$"),
    ("ramp_R_rest_area",    r"^R\s*-\s*Rest Area, Vista Point, Truck Scale$"),
    ("ramp_V_dummy_volume", r"^V\s*-\s*Dummy, Volume only$"),
    ("ramp_Z_other",        r"^Z\s*-\s*Other$"),
]

COLUMN_SPLIT_X = 300  # left column < this, right column >= this
Y_TOLERANCE = 3       # words within this y-distance are same row


# =============================================================================
# PDF -> rows
# =============================================================================

def get_rows_for_column(words, left=True):
    """Group words into (number_or_None, label_text) tuples for one column."""
    col_words = [w for w in words if (w["x0"] < COLUMN_SPLIT_X) == left]
    if not col_words:
        return []
    col_words.sort(key=lambda w: (w["top"], w["x0"]))

    rows = []
    current_top = None
    current = []
    for w in col_words:
        if current_top is None or abs(w["top"] - current_top) <= Y_TOLERANCE:
            current.append(w)
            current_top = w["top"] if current_top is None else current_top
        else:
            rows.append(current)
            current = [w]
            current_top = w["top"]
    if current:
        rows.append(current)

    parsed = []
    for row in rows:
        row.sort(key=lambda w: w["x0"])
        texts = [w["text"] for w in row]
        # Accept thousands separators like "1,918"
        if texts and re.fullmatch(r"-?[\d,]+", texts[0]) and any(c.isdigit() for c in texts[0]):
            num = int(texts[0].replace(",", ""))
            label = " ".join(texts[1:])
        else:
            num = None
            label = " ".join(texts)
        parsed.append((num, label.strip()))
    return parsed


# Noise tokens that get merged into data rows by pdfplumber's row clustering.
# The arrow-decoration + case-insensitive total patterns let the SAME cleaner
# handle the TSN statewide page (whose section headers are bracketed like
# "<------Highway Groups------>" and whose footer is lowercase "Total number of
# Ramps:") as well as the TSMIS per-route page (no brackets, capital "Number").
# Real data labels use a single spaced " - " and never contain "<", ">", or a run
# of 3+ dashes, so stripping those is a no-op on the TSMIS layout (verified).
NOISE_PATTERNS = [
    r"\bHighway Groups\b",
    r"\bOn/Off Indicator\b",
    r"\bPopulation Groups\b",
    r"\bRamp Types\b",
    r"\bNUMBER\s+CODE\b",
    r"\bNUMB\b",
    r"\bCODE\b",
    r"\bER\b",
    r"(?i)\bTotal Number of Ramps:\s*[\d,]+",
    r"(?i)\bRamp Points w/out linework:\s*[\d,]+",
    r"[<>]",          # TSN section-header bracket glyphs
    r"-{3,}",         # TSN section-header dash runs ("------")
]


def clean_label(label):
    """Strip section-header / totals noise that got merged into a data row."""
    out = label
    for pat in NOISE_PATTERNS:
        out = re.sub(pat, "", out)
    return re.sub(r"\s+", " ", out).strip()


def is_new_label(text):
    """Heuristic: does this text start a new schema row, or continue a wrapped one?"""
    if not text:
        return False
    return bool(
        re.match(r"^[A-Z]\s*-\s", text)        # ramp types: "A - ", "D - "
        or re.match(r"^[A-Z]{2,}\s*-\s", text) # on/off:    "ON - ", "OFF - ", "OTH - "
        or re.match(r"^[A-Z]-[A-Z]+", text)    # pop:       "R-RURAL", "U-URBAN"
        or re.match(r"^-[A-Z]", text)          # pop:       "-O", "-I", "-INVALID"
        or text.strip() == "Others"
    )


def _join_continuation(prev, cont):
    """Smart join: 'Co' + 'nnector' -> 'Connector' (broken mid-word),
    but 'Vista Point,' + 'Truck Scale' -> 'Vista Point, Truck Scale'.
    """
    if not prev:
        return cont
    if not cont:
        return prev
    if prev[-1].islower() and cont[0].islower():
        return prev + cont
    return prev + " " + cont


def stitch_wrapped_rows(rows):
    """Combine wrapped continuations and orphan numbers into proper rows."""
    cleaned = [(n, clean_label(l)) for n, l in rows]
    out = []
    open_num, open_label = None, None
    pending_num = None

    def flush():
        nonlocal open_num, open_label
        if open_label is not None or open_num is not None:
            out.append((open_num, open_label or ""))
        open_num, open_label = None, None

    for num, label in cleaned:
        if num is None and not label:
            continue
        if num is not None and not label:
            if open_label is not None and open_num is None:
                open_num = num                  # label was seen first
            else:
                pending_num = num               # hold for next label
            continue
        if num is None and label:
            if is_new_label(label):
                flush()
                open_num = pending_num
                pending_num = None
                open_label = label
            else:
                # continuation of the previous row
                if open_label is not None:
                    open_label = _join_continuation(open_label, label)
                elif out:
                    pn, pl = out[-1]
                    out[-1] = (pn, _join_continuation(pl, label))
            continue
        flush()
        out.append((num, label))
        pending_num = None
    flush()
    return out


def match_schema(rows, schema, used_indices=None):
    """For each schema entry in order, find the next matching row and pull its number."""
    if used_indices is None:
        used_indices = set()
    result = {}
    cursor = 0
    for col_name, pattern in schema:
        found = None
        for i in range(cursor, len(rows)):
            if i in used_indices:
                continue
            num, label = rows[i]
            label_norm = re.sub(r"\s+", " ", label).strip()
            if re.fullmatch(pattern, label_norm):
                found = (i, num)
                break
        if found:
            used_indices.add(found[0])
            cursor = found[0] + 1
            result[col_name] = found[1]
        else:
            result[col_name] = None
    return result


def record_has_data(record):
    """True if a parsed Ramp Summary record carries real ramp figures — not just
    a route read off page 1. A one-page / truncated PDF yields a route-only
    record (the figures live on page 2); such records must not be written as
    blank rows, nor counted as a successful parse."""
    for col_name, value in record.items():
        if col_name in ("source_file", "route") or col_name.startswith("_"):
            continue
        if value is not None:
            return True
    return False


def parse_pdf(path):
    """Return one flat dict for a TSAR ramps summary PDF."""
    record = {"source_file": Path(path).name, "route": None}

    with pdfplumber.open(path) as pdf:
        # Route number from page 1 title
        p1_text = pdf.pages[0].extract_text() or ""
        m = re.search(r"All Ramps on Route\s+(\d+\w*)", p1_text)
        if m:
            record["route"] = m.group(1)

        # Data fields from page 2
        if len(pdf.pages) < 2:
            return record
        page = pdf.pages[1]
        words = page.extract_words()

        left_rows = stitch_wrapped_rows(get_rows_for_column(words, left=True))
        right_rows = stitch_wrapped_rows(get_rows_for_column(words, left=False))

        # Left column: Highway Groups, then On/Off, then Population Groups
        used_left = set()
        record.update(match_schema(left_rows, HIGHWAY_GROUPS, used_left))
        record.update(match_schema(left_rows, ONOFF, used_left))
        record.update(match_schema(left_rows, POP_GROUPS, used_left))

        # Right column: Ramp Types
        record.update(match_schema(right_rows, RAMP_TYPES))

        # Totals are in the page footer, not in either column
        full_text = page.extract_text() or ""
        m = re.search(r"Total Number of Ramps:\s*([\d,]+)", full_text)
        record["total_ramps"] = int(m.group(1).replace(",", "")) if m else None
        m = re.search(r"Ramp Points w/out linework:\s*([\d,]+)", full_text)
        record["ramp_points_no_linework"] = int(m.group(1).replace(",", "")) if m else None

    return record


# =============================================================================
# Records -> Excel
# =============================================================================

# (group_header, [(column_name, display_name)])
GROUPS = [
    ("Source", [
        ("source_file", "Source File"),
        ("route", "Route"),
    ]),
    ("Highway Groups", [(c, c.replace("hwy_", "").title()) for c, _ in HIGHWAY_GROUPS]),
    ("On/Off Indicator", [(c, c.replace("onoff_", "").upper()) for c, _ in ONOFF]),
    ("Population Groups", [
        ("pop_rural_inside",  "Rural-Inside"),
        ("pop_rural_outside", "Rural-Outside"),
        ("pop_urban_inside",  "Urban-Inside"),
        ("pop_urban_outside", "Urban-Outside"),
        ("pop_invalid",       "Invalid"),
    ]),
    ("Ramp Types", [
        ("ramp_A_frontage",    "A-Frontage"),
        ("ramp_B_collector",   "B-Collector"),
        ("ramp_C_connector_L", "C-Conn(L)"),
        ("ramp_D_diamond",     "D-Diamond"),
        ("ramp_E_slip",        "E-Slip"),
        ("ramp_F_connector_R", "F-Conn(R)"),
        ("ramp_G_loop_left",   "G-LoopL"),
        ("ramp_H_buttonhook",  "H-Buttonhook"),
        ("ramp_J_scissors",    "J-Scissors"),
        ("ramp_K_split",       "K-Split"),
        ("ramp_L_loop_noleft", "L-LoopNoL"),
        ("ramp_M_two_way",     "M-TwoWay"),
        ("ramp_P_dummy_paired","P-DummyPair"),
        ("ramp_R_rest_area",   "R-Rest"),
        ("ramp_V_dummy_volume","V-DummyVol"),
        ("ramp_Z_other",       "Z-Other"),
    ]),
    ("Totals", [
        ("total_ramps", "Total Ramps"),
        ("ramp_points_no_linework", "Pts w/o Linework"),
    ]),
    ("Audit", [
        ("_chk_hwy",   "Sum Hwy"),
        ("_chk_onoff", "Sum On/Off + NoLW"),
        ("_chk_pop",   "Sum Pop"),
        ("_chk_ramp",  "Sum RampTypes + NoLW"),
        ("_audit_ok",  "Audit OK"),
    ]),
]

# NOTE: openpyxl style objects (THIN/BORDER/HEADER_FILL) are built inside
# build_workbook, not here, so importing this module never touches openpyxl --
# the GUI can import it even if a dep is missing and get a clean error result.
GROUP_FILLS = {
    "Source":            "4472C4",
    "Highway Groups":    "70AD47",
    "On/Off Indicator":  "ED7D31",
    "Population Groups": "7030A0",
    "Ramp Types":        "C00000",
    "Totals":            "BF8F00",
    "Audit":             "595959",
}

# Long-form labels for the Combined summary sheet — these match the
# code descriptions printed on the source PDF.
LONG_LABELS = {
    "hwy_right":         "R - Right",
    "hwy_divided":       "D - Divided",
    "hwy_undivided":     "U - Undivided",
    "hwy_unconstructed": "X - Unconstructed",
    "hwy_left":          "L - Left",
    "hwy_others":        "Others",
    "onoff_on":          "ON - On",
    "onoff_off":         "OFF - Off",
    "onoff_other":       "OTH - Other",
    "pop_rural_inside":  "R-RURAL -I INSIDE CITY",
    "pop_rural_outside": "       -O OUTSIDE CITY",
    "pop_urban_inside":  "U-URBAN -I INSIDE CITY",
    "pop_urban_outside": "       -O OUTSIDE CITY",
    "pop_invalid":       "       -INVALID DATA",
    "ramp_A_frontage":     "A - Frontage Road",
    "ramp_B_collector":    "B - Collector Road",
    "ramp_C_connector_L":  "C - Direct or Semi-direct Connector (Left)",
    "ramp_D_diamond":      "D - Diamond Type Ramp",
    "ramp_E_slip":         "E - Slip Ramp",
    "ramp_F_connector_R":  "F - Direct or Semi-direct Connector (Right)",
    "ramp_G_loop_left":    "G - Loop (w/Left turn)",
    "ramp_H_buttonhook":   "H - Buttonhook Ramp",
    "ramp_J_scissors":     "J - Scissors",
    "ramp_K_split":        "K - Split Ramp",
    "ramp_L_loop_noleft":  "L - Loop without Left Turn",
    "ramp_M_two_way":      "M - Two way Ramp Segment",
    "ramp_P_dummy_paired": "P - Dummy Paired",
    "ramp_R_rest_area":    "R - Rest Area, Vista Point, Truck Scale",
    "ramp_V_dummy_volume": "V - Dummy, Volume only",
    "ramp_Z_other":        "Z - Other",
}

SUMMARY_SHEET_NAME = "TSAR Ramps Summary"


# --- Combined-sheet layout guard (schema-drift tripwire) -------------------
# build_combined_sheet() hand-places each section at FIXED row anchors: the
# "On/Off Indicator" header at row 13, "Population Groups" at row 19, and the
# Totals at row 28, with data rows auto-filling from each section's first row
# (Highway Groups at 6, On/Off at 15, Population at 21, Ramp Types at 6 in the
# disjoint E-G columns). Those anchors were sized for today's schema lengths.
# If a schema list (HIGHWAY_GROUPS / ONOFF / POP_GROUPS / RAMP_TYPES) GROWS
# beyond its row budget -- enough that its auto-filled rows reach or cross the
# next section's header/Totals anchor -- the Combined sheet is silently
# corrupted, no error raised. This guard turns that overlap into a loud failure:
# grow a section past its budget and you MUST move the next anchor (and the bound
# here) together. (Shrinkage, or growth that still fits the budget, can't
# overrun, so it is intentionally left alone.)
_HG_FIRST_ROW, _ONOFF_HEADER_ROW = 6, 13
_ONOFF_FIRST_ROW, _POP_HEADER_ROW = 15, 19
_POP_FIRST_ROW, _TOTALS_ROW = 21, 28
_RAMP_FIRST_ROW = 6


def _assert_combined_layout():
    """Raise if a section has GROWN past its row budget -- its rows would reach
    or cross the next section's header / Totals anchor and corrupt the Combined
    sheet. A no-op for the current schema (and for shrinkage or in-budget
    growth); a tripwire only for overlap-causing growth."""
    sections = [
        ("Highway Groups", HIGHWAY_GROUPS, _HG_FIRST_ROW, _ONOFF_HEADER_ROW),
        ("On/Off Indicator", ONOFF, _ONOFF_FIRST_ROW, _POP_HEADER_ROW),
        ("Population Groups", POP_GROUPS, _POP_FIRST_ROW, _TOTALS_ROW),
        ("Ramp Types", RAMP_TYPES, _RAMP_FIRST_ROW, _TOTALS_ROW),
    ]
    overruns = []
    for name, schema, first_row, next_anchor in sections:
        last_row = first_row + len(schema) - 1
        if last_row >= next_anchor:
            overruns.append(
                f"{name}: {len(schema)} rows fill {first_row}-{last_row}, "
                f"hitting the anchor at row {next_anchor}")
    if overruns:
        raise ValueError(
            "Ramp Summary Combined-sheet layout drift -- a schema list changed "
            "length but build_combined_sheet's hardcoded row anchors did not: "
            + "; ".join(overruns)
            + ". Move the affected anchor(s) in build_combined_sheet and update "
            "the bounds in _assert_combined_layout together.")


def build_combined_sheet(wb, records, col_letters):
    """Insert a 'Combined' summary sheet at index 0 with formulas that
    aggregate the per-route data on the TSAR Ramps Summary sheet.

    Layout (matches the user-provided template):
      row 1: blue title bar
      row 2: gray subtitle ("Aggregated across N routes...")
      row 4: section headers — "Highway Groups" (left) and "Ramp Types" (right)
      rows 5-11: NUMBER/CODE header + 6 Highway Groups rows
                 and rows 5-21 on the right: 16 Ramp Types rows (incl. P, V)
      rows 13-17: On/Off Indicator section
      rows 19-25: Population Groups section
      rows 28-29: Total Number of Ramps + Ramp Points w/out linework
    The right (Ramp Types) and left (HG/On-Off/Pop) blocks live in disjoint
    columns (E-G vs A-C), so the extended 16-row Ramp Types list cannot collide
    with the Population block even though their row ranges now overlap.
    """
    _assert_combined_layout()      # fail loudly on schema/layout drift (P0 guard)
    ws = wb.create_sheet("Combined", 0)

    n_routes = len(records)
    # Data on the summary sheet starts at row 3. Keep the range valid
    # even when there are zero records so the formulas still evaluate.
    last_data_row = max(3, 2 + n_routes)

    # --- styles (match the uploaded template) ---
    thin = Side(style="thin", color="BFBFBFBF")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=thin, top=thin, bottom=thin)
    section_fill = PatternFill("solid", start_color="0070C0")
    title_fill = PatternFill("solid", start_color="0070C0")
    f_title = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    f_subtitle = Font(name="Arial", color="595959", size=9)
    f_section = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    f_subhead = Font(name="Arial", bold=True, size=10)
    f_data = Font(name="Arial", size=10)
    f_total = Font(name="Arial", bold=True, size=12)
    a_center = Alignment(horizontal="center", vertical="center")
    a_left = Alignment(horizontal="left")
    a_right = Alignment(horizontal="right")

    # --- title + subtitle ---
    ws["A1"] = "All Routes Combined — TSAR Ramp Summary"
    ws.merge_cells("A1:H1")
    ws["A1"].font = f_title
    ws["A1"].fill = title_fill
    ws["A1"].alignment = a_center

    ws["A2"] = (
        f"Aggregated across {n_routes} route(s) — "
        f"live-linked to '{SUMMARY_SHEET_NAME}' sheet"
    )
    ws.merge_cells("A2:H2")
    ws["A2"].font = f_subtitle
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- helpers ---
    def section_header(cell_range, text):
        first = cell_range.split(":")[0]
        ws[first] = text
        ws.merge_cells(cell_range)
        c = ws[first]
        c.font = f_section
        c.fill = section_fill
        c.alignment = a_center
        c.border = section_border

    def subheader_pair(num_cell, code_cell):
        ws[num_cell] = "NUMBER"
        ws[code_cell] = "CODE"
        for addr in (num_cell, code_cell):
            ws[addr].font = f_subhead
            ws[addr].alignment = a_left
            ws[addr].border = box_border

    def data_row(row, num_col, code_range, col_name):
        summary_col = col_letters[col_name]
        num_addr = f"{num_col}{row}"
        ws[num_addr] = (
            f"=SUM('{SUMMARY_SHEET_NAME}'!"
            f"{summary_col}3:{summary_col}{last_data_row})"
        )
        ws[num_addr].font = f_data
        ws[num_addr].alignment = a_right
        ws[num_addr].border = box_border

        first = code_range.split(":")[0]
        ws[first] = LONG_LABELS[col_name]
        ws.merge_cells(code_range)
        ws[first].font = f_data
        ws[first].alignment = a_left
        ws[first].border = box_border

    # --- section headers ---
    section_header("A4:C4", "Highway Groups")
    section_header("E4:G4", "Ramp Types")
    section_header("A13:C13", "On/Off Indicator")
    section_header("A19:C19", "Population Groups")

    # --- NUMBER/CODE sub-headers ---
    subheader_pair("A5", "B5")
    subheader_pair("E5", "F5")
    subheader_pair("A14", "B14")
    subheader_pair("A20", "B20")

    # --- data rows ---
    for i, (col_name, _) in enumerate(HIGHWAY_GROUPS):
        r = 6 + i
        data_row(r, "A", f"B{r}:C{r}", col_name)

    for i, (col_name, _) in enumerate(ONOFF):
        r = 15 + i
        data_row(r, "A", f"B{r}:C{r}", col_name)

    for i, (col_name, _) in enumerate(POP_GROUPS):
        r = 21 + i
        data_row(r, "A", f"B{r}:C{r}", col_name)

    for i, (col_name, _) in enumerate(RAMP_TYPES):
        r = 6 + i
        data_row(r, "E", f"F{r}:G{r}", col_name)

    # --- totals (rows 28-29) ---
    total_col = col_letters["total_ramps"]
    nolw_col = col_letters["ramp_points_no_linework"]
    totals = [
        (28, "Total Number of Ramps:", total_col),
        (29, "Ramp Points w/out linework:", nolw_col),
    ]
    for row, label, summary_col in totals:
        ws[f"A{row}"] = label
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"].font = f_total
        ws[f"A{row}"].alignment = a_right
        ws[f"C{row}"] = (
            f"=SUM('{SUMMARY_SHEET_NAME}'!"
            f"{summary_col}3:{summary_col}{last_data_row})"
        )
        ws[f"C{row}"].font = f_total
        ws[f"C{row}"].alignment = a_left

    # --- column widths + row heights ---
    widths = {"A": 10, "B": 12, "C": 18, "D": 3, "E": 10, "F": 8.43, "G": 38, "H": 3}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    ws.row_dimensions[1].height = 25.5

    # Open the workbook on the Combined sheet by default
    wb.active = wb.index(ws)


def build_workbook(records, out_path):
    """Write a styled, audited workbook with one row per record."""
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HEADER_FILL = PatternFill("solid", start_color="305496")

    wb = Workbook()
    ws = wb.active
    ws.title = "TSAR Ramps Summary"

    # ---- header rows (group + column) ----
    col_idx = 1
    flat_columns = []  # [(col_name, excel_column_index), ...]
    for group_name, cols in GROUPS:
        start = col_idx
        for col_name, display in cols:
            cell = ws.cell(row=2, column=col_idx, value=display)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = BORDER
            flat_columns.append((col_name, col_idx))
            col_idx += 1
        end = col_idx - 1
        ws.cell(row=1, column=start, value=group_name)
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        gcell = ws.cell(row=1, column=start)
        gcell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        gcell.fill = PatternFill("solid", start_color=GROUP_FILLS[group_name])
        gcell.alignment = Alignment(horizontal="center", vertical="center")
        gcell.border = BORDER

    col_letters = {name: get_column_letter(idx) for name, idx in flat_columns}
    col_idx_by_name = dict(flat_columns)

    # ---- data rows ----
    for r, rec in enumerate(records, start=3):
        for col_name, c_idx in flat_columns:
            if col_name.startswith("_chk_") or col_name == "_audit_ok":
                continue  # formulas added below
            v = rec.get(col_name)
            cell = ws.cell(row=r, column=c_idx, value=v)
            if is_formula_injection(v):     # never let parsed text run as a formula
                cell.data_type = "s"
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDER
            if col_name in ("source_file", "route"):
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

        def rng(group_cols):
            letters = [col_letters[c] for c, _ in group_cols]
            return f"{letters[0]}{r}:{letters[-1]}{r}"

        nolw = f"{col_letters['ramp_points_no_linework']}{r}"
        total = f"{col_letters['total_ramps']}{r}"

        formulas = {
            "_chk_hwy":   f"=SUM({rng(HIGHWAY_GROUPS)})",
            "_chk_onoff": f"=SUM({rng(ONOFF)})+{nolw}",
            "_chk_pop":   f"=SUM({rng(POP_GROUPS)})",
            "_chk_ramp":  f"=SUM({rng(RAMP_TYPES)})+{nolw}",
        }
        for col_name, f in formulas.items():
            cell = ws.cell(row=r, column=col_idx_by_name[col_name], value=f)
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

        h, o, pp, rt = (f"{col_letters[c]}{r}" for c in
                        ("_chk_hwy", "_chk_onoff", "_chk_pop", "_chk_ramp"))
        # Self-explaining audit: "OK" when every section's parsed sub-total
        # reconciles to the route's stated Total Number of Ramps, else a message
        # naming the section(s) that don't add up. A red cell here means the
        # SOURCE PDF's own breakdown is short of its own total (a TSMIS data quirk
        # seen on dense routes, e.g. the Ramp Types list omitting a few ramps) --
        # NOT a parsing error -- so the reader knows to check the source, not
        # suspect the tool.
        audit_f = (
            f'=IF(AND({h}={total},{o}={total},{pp}={total},{rt}={total}),"OK",'
            f'"⚠ Source ≠ total: "&'
            f'IF({h}<>{total},"Hwy ","")&IF({o}<>{total},"On/Off ","")&'
            f'IF({pp}<>{total},"Pop ","")&IF({rt}<>{total},"Ramp Types ",""))'
        )
        cell = ws.cell(row=r, column=col_idx_by_name["_audit_ok"], value=audit_f)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---- conditional formatting on audit_ok ----
    if records:
        audit_col_letter = col_letters["_audit_ok"]
        last_row = 2 + len(records)
        audit_range = f"{audit_col_letter}3:{audit_col_letter}{last_row}"
        # Formula-based (text "OK" vs a message) — a cellIs string rule did not
        # apply reliably; EXACT() on the relative first-cell ref adjusts per row.
        first = f"${audit_col_letter}3"
        # CF differential (dxf) fills must set bgColor, not fgColor/start_color
        # (the openpyxl CF gotcha — same pattern compare_core uses for its banner).
        ws.conditional_formatting.add(
            audit_range,
            FormulaRule(formula=[f'EXACT({first},"OK")'],
                        fill=PatternFill(bgColor="C6EFCE"))
        )
        ws.conditional_formatting.add(
            audit_range,
            FormulaRule(formula=[f'NOT(EXACT({first},"OK"))'],
                        fill=PatternFill(bgColor="FFC7CE"))
        )

    # ---- column widths & freeze ----
    width_overrides = {"source_file": 34, "route": 7, "_audit_ok": 30}
    for col_name, c_idx in flat_columns:
        letter = get_column_letter(c_idx)
        if col_name in width_overrides:
            ws.column_dimensions[letter].width = width_overrides[col_name]
        elif col_name.startswith("_chk_"):
            ws.column_dimensions[letter].width = 11
        elif col_name in ("total_ramps", "ramp_points_no_linework"):
            ws.column_dimensions[letter].width = 11
        else:
            ws.column_dimensions[letter].width = 8
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 34
    ws.freeze_panes = "C3"  # freeze headers + first two id columns

    # Combined summary sheet on top, with live formulas referencing
    # the per-route rows on this sheet.
    build_combined_sheet(wb, records, col_letters)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    artifact_store.atomic_save(wb, out_path)        # F9: temp + os.replace (never truncate prior)


# =============================================================================
# Entry point
# =============================================================================

def consolidate(events=None, confirm_overwrite=None, day=None,
                input_dir=None, out_path=None):
    """Parse every per-route Ramp Summary PDF into one audited workbook.

    Console-free: reports progress via events.on_log, asks before overwriting
    through the confirm_overwrite(path)->bool callback, and returns a
    ConsolidateResult. Honors events.is_cancelled() between files.

    `day` picks which export run folder ("<YYYY-MM-DD> <src>-<env>") to read; None means
    the newest run folder, falling back to the legacy flat layout when no run
    folders exist yet.
    """
    events = events or Events()
    if not _DEPS_OK:
        return ConsolidateResult(
            status="error",
            message="Required components are missing (pdfplumber, openpyxl).",
        )
    confirm = confirm_overwrite or (lambda _p: True)
    day = day or latest_output_day()
    input_dir = input_dir or input_dir_for(day)
    out_path = out_path or out_path_for(day)

    if not input_dir.exists():
        return ConsolidateResult(
            status="error",
            message=(f"The {REPORT_NAME} output folder doesn't exist yet:\n{input_dir}\n\n"
                     f"Export the {REPORT_NAME} report first, then consolidate."),
        )

    pdfs = sorted(input_dir.glob("*.pdf"))
    if not pdfs:
        return ConsolidateResult(
            status="error",
            message=(f"No {REPORT_NAME} files were found in:\n{input_dir}\n\n"
                     f"Export the {REPORT_NAME} report first, then consolidate."),
        )

    # Confirm overwrite *before* spending time parsing PDFs.
    if out_path.exists() and not confirm(out_path):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"TSAR Ramp Summary Consolidation - {len(pdfs)} file(s)")
    events.on_log("=" * 60)
    events.on_log("")

    records = []
    failed = []
    blank = []
    for i, p in enumerate(pdfs, 1):
        if events.is_cancelled():
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i:>3}/{len(pdfs)}] {p.name}"
        try:
            rec = parse_pdf(str(p))
        except Exception as e:
            events.on_log(f"{prefix} FAILED ({type(e).__name__}): {e}")
            failed.append(p.name)
            continue
        # A one-page / truncated PDF parses without error but carries no ramp
        # data (page 2 holds the figures) — don't write it as a blank route row,
        # and don't let a folder full of them overwrite a good workbook.
        if record_has_data(rec):
            records.append(rec)
            events.on_log(f"{prefix} parsed")
        else:
            events.on_log(f"{prefix} skipped: no ramp data "
                          "(one-page / truncated PDF?)")
            blank.append(p.name)

    # Nothing usable parsed → do NOT write (a blank/header-only workbook would
    # overwrite a good prior consolidation). Leave the existing file untouched.
    if not records:
        return ConsolidateResult(
            status="error",
            message=(f"None of the {len(pdfs)} {REPORT_NAME} PDF(s) yielded ramp "
                     f"data ({len(failed)} failed to parse, {len(blank)} had no "
                     f"data — truncated/one-page?). Nothing was written and the "
                     f"existing file (if any) was left unchanged.\nRe-export the "
                     f"{REPORT_NAME} report and try again."))

    events.on_log("")
    events.on_log("Writing consolidated workbook...")
    try:
        build_workbook(records, out_path)
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not save {out_path.name}.\n\n"
                     "The file is probably open in Excel. Close it and try again.\n"
                     "(Your exported files were not changed.)"),
        )

    # Loud incomplete banner when anything was left out, so a partial result is
    # never mistaken for a full one.
    incomplete = bool(failed or blank)
    summary_lines = []
    if incomplete:
        summary_lines.append(
            f"⚠ INCOMPLETE — {len(failed) + len(blank)} PDF(s) left OUT "
            f"({len(failed)} failed, {len(blank)} had no data); the workbook "
            f"does NOT contain their routes. Re-export them before relying on it.")
    summary_lines += [
        f"Parsed:      {len(records)}",
        f"Failed:      {len(failed)} {failed if failed else ''}",
        f"No data:     {len(blank)} {blank if blank else ''}",
        f"Output file: {out_path}",
    ]
    return ConsolidateResult(status="ok", output_path=str(out_path),
                             summary_lines=summary_lines,
                             completion=outcome.PARTIAL if incomplete else outcome.COMPLETE,
                             skipped_inputs=len(blank), failed_inputs=len(failed))


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
