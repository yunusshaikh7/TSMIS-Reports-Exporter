"""Build the TSMIS-vs-TSN Highway Summary discrepancy workbook (AGGREGATE).

The Ramp/Intersection Summary recipe applied to the Highway Summary taxonomy, but
measuring MILES instead of counts: each side reduces to ONE statewide
{category: miles} table, compared with has_route=False (key = category, field =
miles). The taxonomy, the codes both systems share, and the familiar-sheet spec
all live in `highway_summary_columns`, shared with the consolidator.

  * TSMIS side — the CONSOLIDATED Highway Summary workbook (one row per route, one
    column per category key); summed column-by-column in exact thousandths.
  * TSN side — the statewide print: a TWO-COLUMN page where each section is its
    own little table. Censused on `Highway Summary Statewide_TSN.pdf` (event
    4843734, reference 09/15/2025).

TSN print conventions, all measured on that print:
  * Each section prints `<---NAME--->` then its own `DVM CODE MILES` header, and
    the sections tile the page in two columns. A section's REGION is its header's
    x-span widened to the midpoints between horizontal neighbours that share a
    y-band, running down to the next section sharing its x-band.
  * Rows are clustered INSIDE a region, never page-wide: the two columns'
    baselines interleave (a right-column row can sit between two left-column
    rows), so a page-wide cluster steals a code line into its neighbour's row and
    orphans its MILES value.
  * The DVM (Daily Vehicle Miles) figure is often GLUED to the code text
    (`6,311,760.663R-RIGHT IND ALIGN`) and is stripped; TSMIS does not tabulate
    DVM, so it is never compared.
  * MEDIAN TYPE labels carry an old-system mapping in parentheses
    (`B-STRIPED (S )`), stripped for matching.
  * RURAL-URBAN prints its `- O - OUTSIDE CITY` rows WITHOUT the parent letter, so
    each binds to the preceding `R-RURAL` / `U-URBAN` row (the CMP-AUD-023 rule);
    a parentless `-O` carrying a value is an error, never defaulted.
  * `Z- NO BARRIER` prints `**********` — the value overflowed its field width.
    That is an ABSENT source fact, not a zero: it is omitted, so the category
    shows one-sided rather than comparing against an invented number.
  * The print's AVERAGE DAILY TRAFFIC section has no TSMIS counterpart and is not
    part of the compared taxonomy.

Console-free; the comparator rides `compare_tsn_common.run_files_compare`.
"""
import re
from dataclasses import replace
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import compare_tsn_common as ctc
import consolidation_meta
import highway_summary_columns as hsc
import summary_layout
from compare_core import CompareSchema
from paths import today_str

REPORT_NAME = "Highway Summary"
TSMIS_SHEET = hsc.SHEET_NAME                      # consolidated per-route sheet
NORMALIZED_SHEET = "Highway Summary (TSN)"        # library normalized 2-col workbook

_SPEC = hsc.summary_spec()
_CATEGORIES = _SPEC.categories()                  # [(key, slug), ...] incl. Total
_KEY_TO_SLUG = {key: slug for key, slug in _CATEGORIES}

# --------------------------------------------------------------------------- #
# TSN print geometry (all measured — see the module docstring)
# --------------------------------------------------------------------------- #
_ROW_TOL = 4.0                 # baseline jitter within one logical row
_DATA_PAGE_RE = re.compile(r"TASAS SELECTIVE RECORD RETRIEVAL", re.I)
_TOTAL_RE = re.compile(r"\*?TOTAL MILES SELECTED\s+([\d,]+\.\d+)", re.I)
_LEAD_DVM_RE = re.compile(r"^[\d,]+\.\d{3}")       # the glued DVM prefix
_PAREN_RE = re.compile(r"\([^)]*\)\s*$")           # the MEDIAN TYPE mapping suffix
_NUM_RE = re.compile(r"^[\d,]+\.\d+$")
_MASK_RE = re.compile(r"^\*+$")                    # an overflowed (masked) value
_LETTER_RE = re.compile(r"^([A-Z+])\s*-")
# The TSN-only section; present in the print, absent from the TSMIS export.
TSN_ONLY_SECTIONS = ("AVERAGE DAILY TRAFFIC",)


def _rows_of(words, tol=_ROW_TOL):
    out = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        for r in out:
            if abs(r["top"] - w["top"]) <= tol:
                r["w"].append(w)
                break
        else:
            out.append({"top": w["top"], "w": [w]})
    for r in out:
        r["w"].sort(key=lambda w: w["x0"])
    return sorted(out, key=lambda r: r["top"])


def _regions(words, page_h):
    """Every `<---NAME--->` section on the page, as a parsing REGION."""
    secs, i = [], 0
    while i < len(words):
        if "<-" in words[i]["text"]:
            j = i
            while j < len(words) and "->" not in words[j]["text"]:
                j += 1
            if j < len(words):
                txt = " ".join(w["text"] for w in words[i:j + 1])
                name = re.sub(r"^<-+\s*|\s*-+>$", "", txt).strip(" -")
                secs.append({"name": hsc.norm_label(name), "top": words[i]["top"],
                             "x0": words[i]["x0"], "x1": words[j]["x1"]})
                i = j
        i += 1
    for s in secs:                       # down to the next section sharing an x-band
        nxt = [o["top"] for o in secs if o["top"] > s["top"] + 1
               and not (o["x1"] < s["x0"] or o["x0"] > s["x1"])]
        s["y_end"] = min(nxt) if nxt else page_h
    for s in secs:                       # out to the midpoints of its y-band neighbours
        def shares_y(o, s=s):
            return not (o["y_end"] <= s["top"] or o["top"] >= s["y_end"])
        rights = [o["x0"] for o in secs if o is not s and o["x0"] > s["x1"] and shares_y(o)]
        lefts = [o["x1"] for o in secs if o is not s and o["x1"] < s["x0"] and shares_y(o)]
        s["left"] = (max(lefts) + s["x0"]) / 2 if lefts else 0.0
        s["right"] = (s["x1"] + min(rights)) / 2 if rights else float("inf")
    return secs


def _tsn_code(section, text, parent, source):
    """The within-section code for a TSN row's label, plus the updated RURAL-URBAN
    parent. Mirrors `highway_summary_columns._code_for` for the TSMIS spelling."""
    up = hsc.norm_label(text)
    if section == "NON-ADD":
        return "N", parent
    if section == "RURAL-URBAN":
        if up.startswith("--"):                       # '--INVALID DATA'
            return "INVALID", parent
        if up.startswith("-"):                        # '- O - OUTSIDE CITY'
            if parent is None:
                raise ValueError(
                    f"{source}: a RURAL-URBAN '- O - OUTSIDE CITY' row has no "
                    "preceding R-RURAL/U-URBAN parent — its mileage cannot be "
                    "attributed")
            return f"{parent}-O", parent
        return up[0], up[0]                           # R-RURAL… / U-URBAN…
    if section in ("MEDIAN WIDTH", "NUMBER OF LANES", "DESIGN SPEED",
                   "AVERAGE DAILY TRAFFIC"):
        code = hsc.range_code(text)
        return (code if code is not None else up), parent
    m = _LETTER_RE.match(up)
    return (m.group(1) if m else up), parent


def parse_tsn_pdf(path):
    """The statewide TSN Highway Summary print -> {slug: miles-thousandths}.

    Only categories the print actually carries a NUMBER for are returned: a
    masked (`**********`) value is omitted, never coerced to 0, so the comparator
    shows it one-sided instead of comparing an invented figure (CMP-AUD-021)."""
    index = hsc.code_index("tsn")
    out, total, seen = {}, None, set()
    name = Path(path).name
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if not _DATA_PAGE_RE.search(text):
                continue
            if total is None:
                m = _TOTAL_RE.search(text)
                if m:
                    total = hsc.parse_miles(m.group(1), source=name,
                                            category=hsc.TOTAL_LABEL)
            words = page.extract_words()
            for s in _regions(words, page.height):
                if s["name"] in (hsc.norm_label(n) for n in TSN_ONLY_SECTIONS):
                    continue
                region = [w for w in words
                          if s["left"] <= w["x0"] < s["right"]
                          and s["top"] + 1 < w["top"] < s["y_end"]]
                parent = None
                for r in _rows_of(region):
                    toks = r["w"]
                    if len(toks) < 2:
                        continue
                    last = toks[-1]["text"]
                    masked = bool(_MASK_RE.match(last))
                    if not (masked or _NUM_RE.match(last)):
                        continue
                    label = " ".join(w["text"] for w in toks[:-1])
                    label = _PAREN_RE.sub("", _LEAD_DVM_RE.sub("", label).strip()).strip()
                    if not label:
                        continue
                    code, parent = _tsn_code(s["name"], label, parent, name)
                    cat = index.get((s["name"], code))
                    if cat is None:
                        continue          # a section/code the taxonomy doesn't share
                    if cat.slug in seen:
                        raise ValueError(
                            f"{name}: the {s['name']} category {label!r} appears "
                            "twice — refusing to read an ambiguous table")
                    seen.add(cat.slug)
                    if masked:
                        continue          # absent, not zero
                    out[cat.slug] = hsc.parse_miles(last, source=name,
                                                    category=cat.key)
    if total is None:
        raise ValueError(f"{name}: no 'TOTAL MILES SELECTED' line was found — this "
                         "does not look like a TSN Highway Summary print")
    out[hsc.TOTAL_SLUG] = total
    return out


def suggest_name(tsmis_path):
    return f"TSMIS_vs_TSN_HighwaySummary_Comparison {today_str()}.xlsx"


def parse_tsn_source_claims(path):
    """The print's source claims (CMP-AUD-146) — identity/timing/submitter."""
    path = Path(path)
    try:
        with pdfplumber.open(path) as pdf:
            full_text = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
    except Exception as e:
        raise ValueError(f"Could not read {path.name}: {type(e).__name__}: {e}")
    return {"schema_version": 1,
            "identity": ctc.tsn_print_identity(full_text, path.name)}


def claims_notes(claims, side_label="TSN"):
    """Human-readable exposure lines for the familiar sheet + log."""
    if not claims:
        return [f"{side_label} print: no source-claims record beside this "
                "normalized workbook (older normalization) — rebuild the TSN "
                "library to capture the print identity."]
    ident = claims.get("identity") or {}
    if not ident:
        return []
    return [f"{side_label} print identity: {ident.get('report_id')} · Event "
            f"{ident.get('event_id')} · reference {ident.get('reference_date')} "
            f"· submitted by {ident.get('submitter')} · generated "
            f"{ident.get('generated_time')} ({ident.get('location_criteria')})."]


# --------------------------------------------------------------------------- #
# loaders
# --------------------------------------------------------------------------- #
def _load_tsn(path):
    """TSN side -> {slug: milli}. Reads the raw statewide print, or the library's
    normalized 2-column workbook (Category | Miles) if that was supplied."""
    path = Path(path)
    if path.suffix.lower() == ".pdf":
        try:
            return parse_tsn_pdf(str(path))
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"Could not read {path.name}: {type(e).__name__}: {e}")
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[NORMALIZED_SHEET] if NORMALIZED_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
        out = {}
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None or str(header[0]).strip().casefold() != "category":
            raise ValueError(
                f"{path.name}: this is not a normalized TSN Highway Summary "
                "workbook (expected a 'Category' | 'Miles' header)")
        for r in rows:
            if not r or r[0] is None:
                continue
            slug = _KEY_TO_SLUG.get(str(r[0]).strip())
            if slug is None or r[1] is None:
                continue
            out[slug] = hsc.parse_miles(r[1], source=path.name, category=str(r[0]))
        return out
    finally:
        wb.close()


def _load_tsmis(path):
    """TSMIS side -> {slug: milli}: the CONSOLIDATED workbook's per-route sheet,
    summed column by column in exact thousandths."""
    path = Path(path)
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        if TSMIS_SHEET not in wb.sheetnames:
            raise ValueError(
                f"{path.name}: no '{TSMIS_SHEET}' sheet — pick the CONSOLIDATED "
                "Highway Summary workbook (Consolidate ▸ Highway Summary).")
        ws = wb[TSMIS_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = [("" if h is None else str(h).strip()) for h in (next(rows, None) or [])]
        if hsc.recognize(header) is None:
            raise ValueError(
                f"{path.name}: the '{TSMIS_SHEET}' sheet does not carry the "
                "Highway Summary layout — re-consolidate it with this app version.")
        col = {h: i for i, h in enumerate(header)}
        out = {slug: 0 for _key, slug in _CATEGORIES}
        routes = 0
        for r in rows:
            if not r or r[0] is None or str(r[0]).strip() == "":
                continue
            routes += 1
            for key, slug in _CATEGORIES:
                v = r[col[key]] if col[key] < len(r) else None
                if v is None:
                    continue
                out[slug] += hsc.parse_miles(v, source=path.name, category=key)
        if not routes:
            raise ValueError(f"{path.name}: the consolidated workbook has no routes.")
        return out
    finally:
        wb.close()


def _rows(values):
    """{slug: milli} -> the compared [category, miles] rows, in spec order. A slug
    with no value is OMITTED so it shows one-sided instead of comparing a zero
    the source never stated."""
    return [[key, hsc.miles(values[slug])]
            for key, slug in _CATEGORIES if values.get(slug) is not None]


def _load_pair(tsmis_path, tsn_path, note_sink=None, events=None):
    tsmis = _load_tsmis(tsmis_path)
    tsn = _load_tsn(tsn_path)
    notes = []
    claims = None
    if Path(tsn_path).suffix.lower() == ".pdf":
        claims = parse_tsn_source_claims(tsn_path)
    else:
        claims = consolidation_meta.read_extra(tsn_path, "tsn_source_claims")
    notes += claims_notes(claims)
    missing = [key for key, slug in _CATEGORIES
               if slug not in tsn and any(c.code and c.sides in ("both", "tsn")
                                          for c in hsc.CATS if c.slug == slug)]
    if missing:
        notes.append(
            "TSN print did not state a mileage for "
            + ", ".join(missing[:4]) + ("…" if len(missing) > 4 else "")
            + " (the print masks a value that overflows its column width with "
            "'**********'); shown one-sided rather than compared as zero.")
    if note_sink is not None:
        note_sink.clear()
        note_sink.extend(notes)
    if events is not None:
        for n in notes:
            events.on_log(f"note: {n}")
    return _rows(tsmis), _rows(tsn), []


_SCHEMA = CompareSchema(
    report_name=REPORT_NAME,
    header=["Category", "Miles"],
    side_a="TSMIS",
    side_b="TSN",
    id_noun="category",
    id_noun_plural="categories",
    sides_noun="systems",
    cmp_widths={"Miles": 14},
    data_widths={"Miles": 14},
    scope_flat="Statewide category mileage",
    one_sided_note_extra=" (a category one system classifies and the other doesn't)",
    extra_sheet_writer=summary_layout.make_extra_sheet_writer(_SPEC),
)


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Highway Summary TSMIS-vs-TSN AGGREGATE comparison workbook(s)."""
    notes = []
    schema = replace(_SCHEMA, extra_sheet_writer=summary_layout.make_extra_sheet_writer(
        _SPEC, extra_notes=notes))
    return ctc.run_files_compare(
        schema, tsmis_path, tsn_path, out_path,
        banner="Highway Summary Comparison — TSMIS vs TSN (statewide category miles)",
        has_route=False,
        loader=lambda a, b: _load_pair(a, b, note_sink=notes, events=events),
        deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (pdfplumber, openpyxl).",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
