"""Single source of truth for the Highway Log column labels + meanings.

The vendor TSMIS Excel export MISLABELS most Highway Log columns — "N/A" is
actually Non-Add Mileage (not "not applicable"), every roadbed column is a
cryptic 1-2 letter code ("LB T" is the Left-roadbed Surface Type, "LB OT" is the
outside-shoulder TOTAL width…), and it even labels two different columns "RB SH".
Those wrong labels were taken as fact and propagated to every Highway Log
workflow (the TSN converter, the consolidators, the comparisons).

The CORRECT meanings come from the report's OWN legend (the TSMIS legend page /
the TSN OTM52010 header), verified against the printed document. This module is
the ONE place that defines them, so every Highway Log workflow agrees:

  * HEADER         the 31 canonical labels (doc abbreviation + the vendor's old
                   label in [brackets] where it was wrong), in fixed column order
  * VENDOR_HEADER  the 31 OLD vendor labels — so a workbook exported/built before
                   this overhaul is still RECOGNIZED (compared by position) and
                   relabeled to the canonical header
  * COLUMNS        (group, label, vendor, meaning) for the Legend sheet + tooltips

Column order is the TSMIS layout: Location & Distance, general, LEFT ROADBED,
MEDIAN, RIGHT ROADBED, Description, dates. (The TSN log additionally carries an
ADT Information group — Look Back / P / Look Ahead — which the TSMIS format drops,
so it is not part of this header.) Description is column 28: the PDF prints it on
its own line below each segment, not as a header column in the data table.

Kept import-light (the data + recognize() need no third-party libs); the
openpyxl-based Legend/tooltip helpers guard their import so importing this module
never fails when a dependency is missing.
"""
import re

# (group, canonical label, exact vendor label, plain-English meaning)
# group "" = an ungrouped column (no roadbed/median/location band over it).
COLUMNS = [
    ("Location & Distance", "Location", "Location",
     "Postmile (PM) — the route location of the segment"),
    ("Location & Distance", "Length (MI) [MI]", "MI",
     "Segment length in miles (the distance to the next postmile)"),
    ("Location & Distance", "NA [N/A]", "N/A",
     "Non-Add Mileage — NOT 'not applicable'"),
    ("Location & Distance", "Cnty Odom", "Cnty Odom",
     "County odometer (the county postmile / odometer reading)"),
    ("Location & Distance", "City", "City", "City code"),
    ("", "RU [R/U]", "R/U", "Rural / Urban / Urbanized"),
    ("", "SPD", "SPD", "Design Speed"),
    ("", "TER", "TER", "Terrain"),
    ("", "HG [H/G]", "H/G", "Highway Group"),
    ("", "AC [A/C]", "A/C", "Access Control"),
    ("Left Roadbed", "LB ST [LB T]", "LB T",
     "Left roadbed — Surface Type (ST)"),
    ("Left Roadbed", "LB # Lns [LB Lns]", "LB Lns",
     "Left roadbed — Number of Lanes (# Lns)"),
    ("Left Roadbed", "LB SF [LB F]", "LB F",
     "Left roadbed — Special Features (SF)"),
    ("Left Roadbed", "LB OT-SH Total [LB OT]", "LB OT",
     "Left roadbed — Outside Shoulder (OT-SH), Total width (TO)"),
    ("Left Roadbed", "LB OT-SH Treated [LB TR]", "LB TR",
     "Left roadbed — Outside Shoulder (OT-SH), Treated width (TR)"),
    ("Left Roadbed", "LB T-W Wid [LB T-W]", "LB T-W",
     "Left roadbed — Traveled Way Width (T-W Wid)"),
    ("Left Roadbed", "LB IN-SH Total [LB IN]", "LB IN",
     "Left roadbed — Inside Shoulder (IN-SH), Total width (TO)"),
    ("Left Roadbed", "LB IN-SH Treated [LB SH]", "LB SH",
     "Left roadbed — Inside Shoulder (IN-SH), Treated width (TR)"),
    ("Median", "Med TY/CL/BA [Med TCB]", "Med TCB",
     "Median — Type (TY) / Curb & Landscape (CL) / Barrier (BA)"),
    ("Median", "Med Wid/Var [Med Wid]", "Med Wid",
     "Median — Width (Wid) / Variance (Var)"),
    ("Right Roadbed", "RB ST [RB T]", "RB T",
     "Right roadbed — Surface Type (ST)"),
    ("Right Roadbed", "RB # Lns [RB Lns]", "RB Lns",
     "Right roadbed — Number of Lanes (# Lns)"),
    ("Right Roadbed", "RB SF [RB F]", "RB F",
     "Right roadbed — Special Features (SF)"),
    ("Right Roadbed", "RB IN-SH Total [RB IN]", "RB IN",
     "Right roadbed — Inside Shoulder (IN-SH), Total width (TO)"),
    ("Right Roadbed", "RB IN-SH Treated [RB SH]", "RB SH",
     "Right roadbed — Inside Shoulder (IN-SH), Treated width (TR) "
     "(the vendor labeled this 'RB SH', same as the outside-shoulder column)"),
    ("Right Roadbed", "RB T-W Wid [RB T-W]", "RB T-W",
     "Right roadbed — Traveled Way Width (T-W Wid)"),
    ("Right Roadbed", "RB OT-SH Total [RB OT]", "RB OT",
     "Right roadbed — Outside Shoulder (OT-SH), Total width (TO)"),
    ("Right Roadbed", "RB OT-SH Treated [RB SH]", "RB SH",
     "Right roadbed — Outside Shoulder (OT-SH), Treated width (TR) "
     "(the vendor labeled this 'RB SH', same as the inside-shoulder column)"),
    ("", "Description", "Description",
     "Feature description (the TSMIS/TSN log prints it on its own line below "
     "each segment)"),
    ("", "Date of Rec", "Date of Rec", "Date of Record"),
    ("", "Sig Chg. Date", "Sig Chg. Date", "Significant Change Date"),
]

HEADER = [c[1] for c in COLUMNS]            # the 31 canonical labels
VENDOR_HEADER = [c[2] for c in COLUMNS]     # the 31 old vendor labels (recognized too)
DESC_IDX = HEADER.index("Description")      # 28 — filled from follow-on PDF lines
ROUTE_COL = "Route"                         # leading column on consolidated workbooks

assert DESC_IDX == 28 and len(HEADER) == 31     # layout guard
assert len(VENDOR_HEADER) == 31

_ACCEPTED = (HEADER, VENDOR_HEADER)


def recognize(header):
    """Is `header` (a loaded row-1 list, possibly with a leading 'Route') a
    Highway Log layout — the canonical labels OR the old vendor labels?
    Returns has_route (bool) when recognized, else None. Comparison is by the
    full label list, so a same-width but different report can't sneak through;
    accepting both label sets means a workbook built before this overhaul still
    compares (the engine aligns by column POSITION and relabels to HEADER)."""
    for base in _ACCEPTED:
        if header == list(base):
            return False
        if header == [ROUTE_COL] + list(base):
            return True
    return None


# ---------------------------------------------------------------------------
# Ditto markers (`+` / `++` / `+++`) — the "see paired roadbed" convention.
# Validated structurally against 110k+ real rows (TSMIS PDF + TSN): a ditto cell
# means "this attribute is not the subject of this row; its value is the one given
# on the PAIRED roadbed's own row." A Right-roadbed row dittos its Left block; a
# Left-roadbed row its Right block; combined rows ditto neither.
#
# It overwhelmingly fills a FULLY-dittoed roadbed block (the 8 Left or 8 Right
# columns), but it is NOT confined to those two blocks: on divided-highway rows
# the SHARED median/access-control attributes (Access Control + the two Median
# columns) are dittoed too, alongside the dittoed roadbed block — these sit
# OUTSIDE the roadbed blocks (the 2026-06-18 audit found ~hundreds of such cells
# on the TSN side). So both the non-asserting diff rule AND this display fill are
# COLUMN-AGNOSTIC (keyed on the `+`-run shape, not the column), to catch a ditto
# wherever a source places it.
#
# A ditto is a POINTER, not data — so in a comparison it is NON-ASSERTING (it
# never counts as a difference; see compare_core._is_plus_run, also column-
# agnostic). The paired value is filled only for DISPLAY.
# See docs/highway_log/comparison-study.md.
# ---------------------------------------------------------------------------

# Column-index positions (within HEADER) of the two 8-column roadbed blocks.
LEFT_BLOCK_IDX = tuple(i for i, c in enumerate(COLUMNS) if c[0] == "Left Roadbed")
RIGHT_BLOCK_IDX = tuple(i for i, c in enumerate(COLUMNS) if c[0] == "Right Roadbed")
assert len(LEFT_BLOCK_IDX) == 8 and len(RIGHT_BLOCK_IDX) == 8     # layout guard

_DITTO_RE = re.compile(r"\++")            # one or more '+': '+', '++', '+++'
_PM_RE = re.compile(r"\d{3}\.\d{3}")      # the bare postmile inside a Location


def is_ditto(value):
    """True if `value` is a ditto marker — a run of one or more '+' ('+', '++',
    '+++'). A dittoed roadbed-block cell points at the paired roadbed's row and
    is NOT a data value."""
    if value is None:
        return False
    s = str(value).strip()
    return bool(s) and _DITTO_RE.fullmatch(s) is not None


def _base_postmile(location):
    """The bare 'ddd.ddd' postmile inside a Location token (drops the optional
    realignment prefix and roadbed/equation suffix), used to pair a Right-roadbed
    row with the Left-roadbed row at the SAME stretch."""
    m = _PM_RE.search(str(location or ""))
    return m.group(0) if m else str(location or "")


def fill_paired_roadbed(rows, loc_idx=0):
    """DISPLAY-ONLY ditto fill for ONE route's rows (each a list aligned to
    HEADER, Location at `loc_idx`, in document order).

    Returns (filled_rows, ditto_cells): a deep copy of `rows` with every dittoed
    cell replaced by the paired roadbed's value, and a set of (row_index,
    col_index) marking which cells were ditto-derived (so the caller can flag
    them). Scanning is COLUMN-AGNOSTIC — a ditto is detected by its `+`-run shape
    in ANY column, not only the two roadbed blocks, so the shared median/access-
    control dittos on divided-highway rows are caught too (matching the column-
    agnostic non-asserting diff rule). The fill source for a dittoed cell is the
    same column on the nearest row whose value there is concrete, PREFERRING a row
    at the same base postmile (the true paired roadbed); an unfillable ditto is
    left as-is and still marked. Because the comparison treats ditto as
    non-asserting, this fill never affects a diff result — it is purely
    informational."""
    filled = [list(r) for r in rows]
    ditto_cells = set()
    n = len(rows)
    width = len(HEADER)

    def cell(j, col):
        return rows[j][col] if col < len(rows[j]) else None

    for i in range(n):
        base_i = _base_postmile(cell(i, loc_idx))
        for col in range(width):
            if col == loc_idx or not is_ditto(cell(i, col)):
                continue
            ditto_cells.add((i, col))
            same_base, nearest = None, None
            for j in range(n):
                if j == i:
                    continue
                v = cell(j, col)
                if v is None or is_ditto(v) or str(v).strip() == "":
                    continue
                if _base_postmile(cell(j, loc_idx)) == base_i:
                    same_base = v
                    break                       # exact paired roadbed wins
                d = abs(j - i)
                if nearest is None or d < nearest[0]:
                    nearest = (d, v)
            value = same_base if same_base is not None else (
                nearest[1] if nearest else None)
            if value is not None and col < len(filled[i]):
                filled[i][col] = value
    return filled, ditto_cells


def display_fills(rows, has_route):
    """Comparison DISPLAY helper: {row_index: {col_in_row: resolved_value}} for
    every dittoed roadbed cell, computed per route via fill_paired_roadbed.

    `rows` are the comparison's input rows (each a list; a leading Route column
    when has_route). `row_index` is the 0-based index into `rows`; `col_in_row`
    is the cell's position in that row (the roadbed-block HEADER index, shifted by
    the leading Route). The caller (the data-sheet writer) attaches the resolved
    value as a cell comment + tint, so a reviewer SEES what a `+`/`++` resolved to
    — while the cell itself keeps the raw ditto (the non-asserting diff depends on
    it). The resolution is display-only and never affects a diff result."""
    off = 1 if has_route else 0
    groups = {}                          # route -> [(global_index, row), ...]
    for gi, r in enumerate(rows):
        route = r[0] if has_route else ""
        groups.setdefault(route, []).append((gi, r))
    out = {}
    for items in groups.values():
        sub = [r[off:] for _, r in items]            # align to HEADER (Location at 0)
        filled, ditto_cells = fill_paired_roadbed(sub, loc_idx=0)
        for local_i, col in ditto_cells:
            gi = items[local_i][0]
            out.setdefault(gi, {})[col + off] = filled[local_i][col]
    return out


# ---------------------------------------------------------------------------
# Roadbed-aware comparison key. TSMIS (PDF + Excel) tag a divided segment's
# roadbed with a trailing 'R'/'L' in the Location ('R021.466R'); TSN omits the
# suffix ('R021.466') and encodes the roadbed implicitly via which 8-col block
# is dittoed. Keying on the raw Location therefore SPLITS the same physical
# roadbed row into a false one-sided pair across a TSMIS-vs-TSN comparison. The
# 2026-06-18 audit proved the suffix ↔ dittoed-block correspondence is 100%
# (suffix 'R' ⇔ Left block dittoed ⇔ right roadbed), so the roadbed identity is
# recoverable on BOTH sides — this canonical key unifies the two encodings.
# Opt-in via CompareSchema.key_normalizer (only the TSMIS-vs-TSN schemas set it;
# same-encoding comparisons — cross-env TSMIS-vs-TSMIS, PDF-vs-Excel — are
# unaffected). See docs/highway_log/comparison-study.md §7b.
# ---------------------------------------------------------------------------

def roadbed_tag(row, off=0):
    """The roadbed a row describes — 'R' (right), 'L' (left), or '' (combined /
    indeterminate) — derived from which 8-col block is dittoed. A Left-block-
    dittoed row describes the RIGHT roadbed (its Left geometry points at the
    paired row), and vice-versa."""
    ld = sum(1 for c in LEFT_BLOCK_IDX
             if off + c < len(row) and is_ditto(row[off + c]))
    rd = sum(1 for c in RIGHT_BLOCK_IDX
             if off + c < len(row) and is_ditto(row[off + c]))
    if ld and not rd:
        return "R"
    if rd and not ld:
        return "L"
    return ""


def roadbed_canonical_location(row, off=0, key_field=0):
    """Canonical roadbed-aware Location key for a Highway Log row, unifying the
    PDF/Excel suffix encoding with TSN's dittoed-block encoding so the SAME
    physical roadbed row keys identically on both sides.

    Returns <base><roadbed-letter>:
      * a Location already ending in a roadbed 'R'/'L' (PDF/Excel) is authoritative
        and returned unchanged;
      * a suffix-less Location (TSN) gets the block-derived roadbed tag appended.
    The trailing equation marker 'E' and the leading alignment prefix are PRESERVED
    (they are identity, not the roadbed) — so a genuine route-start 'R000.000'
    never collapses into a bridge '000.000', and an equation 'E' variant is not
    merged. `off` is the leading-Route offset (1 on consolidated rows); `key_field`
    is the Location column index in the non-Route header (0 for the Highway Log)."""
    li = off + key_field
    s = "" if li >= len(row) or row[li] is None else str(row[li]).strip()
    if s and s[-1] in ("R", "L"):
        return s
    return s + roadbed_tag(row, off)


def legend_rows():
    """Rows for the 'Legend' sheet: (group, canonical label, vendor label,
    meaning), in column order. 'group' blank for ungrouped columns."""
    return [(grp, label, vendor, meaning) for grp, label, vendor, meaning in COLUMNS]


def tooltip_for(label):
    """The hover-tooltip text for a header `label` (group prefix + meaning), or
    None when the label isn't a Highway Log column."""
    for grp, lab, _vendor, meaning in COLUMNS:
        if lab == label:
            return f"[{grp}] {meaning}" if grp else meaning
    return None


# ---------------------------------------------------------------------------
# openpyxl helpers (guarded — importing this module must never need openpyxl).
# Used by the consolidators and the comparison engine to attach a hover tooltip
# to every Highway Log header cell and to drop a "Legend" sheet into the book.
# Work in both normal and write_only (streaming) workbooks.
# ---------------------------------------------------------------------------
try:
    from openpyxl.comments import Comment
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPX = True
except ImportError:
    _OPX = False

_LEGEND_DARK = "1F3864"
_LEGEND_TITLE = ("Highway Log — column legend (corrected labels). The vendor "
                 "Excel export mislabeled these; the bracketed [old] label is "
                 "what it used.")


def comment_for(label):
    """An openpyxl Comment (hover tooltip) for header `label`, or None."""
    if not _OPX:
        return None
    text = tooltip_for(label)
    if text is None:
        return None
    c = Comment(text, "TSMIS Exporter")
    c.width, c.height = 320, 120
    return c


def write_legend_sheet(wb, title="Legend"):
    """Append a Legend worksheet (Group / Column / Vendor label (old) / Meaning)
    to `wb`. Streaming-safe: only create_sheet + append are used, so it works on
    a write_only workbook too. No-op if openpyxl is unavailable."""
    if not _OPX:
        return None
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = _LEGEND_DARK
    hfont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color=_LEGEND_DARK)
    note_font = Font(name="Arial", size=10, italic=True, color="595959")
    body = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    write_only = getattr(wb, "write_only", False)

    def cell(value, font=body, fill=None, align=None):
        if write_only:
            c = WriteOnlyCell(ws, value=value)
            c.font = font
            if fill:
                c.fill = fill
            if align:
                c.alignment = align
            return c
        return value

    for col, w in (("A", 18), ("B", 26), ("C", 16), ("D", 64)):
        ws.column_dimensions[col].width = w
    ws.append([cell(_LEGEND_TITLE, note_font)])
    ws.append([cell(h, hfont, hfill) for h in
               ("Group", "Column", "Vendor label (old)", "Meaning")])
    for grp, label, vendor, meaning in legend_rows():
        same = (label == vendor)
        ws.append([cell(grp or "—"), cell(label),
                   cell("" if same else vendor),
                   cell(meaning, body, align=wrap)])
    ws.append([])
    ws.append([cell("Ditto", body),
               cell("+ / ++ / +++", body),
               cell("", body),
               cell("A column printed as '+' marks means 'same as the PAIRED "
                    "roadbed's own row' (a Right-roadbed row dittos its Left "
                    "block, and vice-versa; on divided highways the shared "
                    "median/access-control columns ditto too) — it is a pointer, "
                    "not data. In a comparison these are NEVER counted as a "
                    "difference; a tinted ditto cell's hover note shows the value "
                    "it resolves to.", body, align=wrap)])
    if not write_only:
        for row in ws.iter_rows(min_row=2, max_row=2):
            for c in row:
                c.font = hfont
                c.fill = hfill
    return ws


def apply_header_tooltips(ws, header_cells=None, first_row=1):
    """Attach the hover tooltip to each Highway Log header cell on an
    already-written, NON-streaming sheet (random-access). For write_only sheets
    the caller sets `.comment` on the WriteOnlyCell BEFORE appending instead
    (see comment_for). No-op if openpyxl is unavailable."""
    if not _OPX:
        return
    for c in next(ws.iter_rows(min_row=first_row, max_row=first_row)):
        cm = comment_for(c.value)
        if cm is not None:
            c.comment = cm
