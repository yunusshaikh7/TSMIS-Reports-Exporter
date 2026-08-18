"""Single source of truth for the Highway Detail column labels + meanings.

Unlike the Highway Log, the TSMIS Highway Detail Excel export labels its 34
columns CORRECTLY (they match the report's own legend page), so no relabel
override is needed — this module exists so every Highway Detail workflow (the
Excel consolidator, the PDF-sourced consolidator, the comparisons) agrees on the
one header, and so header cells can carry the legend meanings as hover tooltips.

The meanings come from the report's OWN legend (the TSMIS legend page, verified
against the TSN OTM22220 "TSAR-HIGHWAY DETAIL" legend): each record is two
printed lines — line 1 the post mile / length / record / access-control / city
attributes with their effective dates, line 2 the description and the Left
Roadbed / Median / Right Roadbed attribute blocks, each led by its own
effective date. The legacy ADT Information block is TSN-only (the TSMIS report
intentionally omits it), so it is not part of this header.

Kept import-light (the data + recognize() need no third-party libs); the
openpyxl-based Legend/tooltip helpers guard their import so importing this
module never fails when a dependency is missing.
"""

# (group, label, plain-English meaning) — label == the export's own header text.
# group "" = an ungrouped column (no roadbed/median band over it).
COLUMNS = [
    ("", "Post Mile",
     "Postmile: prefix + mile + marker (e.g. 'R012.243', 'S000.000', '000.000E'). "
     "Prefix codes: C commercial lanes, D duplicate PM at a meandering county "
     "line, G reposting at a route end, H realignment of D, L overlap, M "
     "realignment of R, N realignment of M, R first realignment, S spur, T "
     "temporary connection. A trailing R/L marks a Right/Left independent-"
     "alignment roadbed row; a trailing E marks an equation point."),
    ("", "Length", "Distance to the next record (miles)"),
    ("", "Date of Rec", "Network date of record"),
    ("", "HG", "Highway Group: R/L independent alignment, D divided, "
               "U undivided, X unconstructed"),
    ("", "AC", "Access Control"),
    ("", "Acc-Cont Eff", "Access Control effective date"),
    ("", "City", "City code"),
    ("", "RU", "Rural / Urban (Population Code)"),
    ("", "RU Eff", "Rural/Urban (Population Code) effective date. NOTE: the "
                   "legacy TASAS/TSN report prints the ADT profile BEGIN date "
                   "in this slot, so it differs from TSN on nearly every row."),
    ("", "Description", "Feature description"),
    ("", "NA", "Non-Add Mileage ('N' when the mileage is non-add; blank "
               "otherwise — TSN prints an explicit 'A' for add mileage)"),
    ("Left Roadbed", "LB Eff", "Left roadbed — section effective date"),
    ("Left Roadbed", "LB S/T", "Left roadbed — Surface Type"),
    ("Left Roadbed", "LB #Ln", "Left roadbed — Number of Lanes"),
    ("Left Roadbed", "LB S/F", "Left roadbed — Special Feature ('Z' = none; "
                               "TSMIS leaves this blank where TSN prints Z)"),
    ("Left Roadbed", "LB OT-TO", "Left roadbed — Outside Shoulder, Total width"),
    ("Left Roadbed", "LB OT-TR", "Left roadbed — Outside Shoulder, Treated width"),
    ("Left Roadbed", "LB Wid", "Left roadbed — Traveled Way Width"),
    ("Left Roadbed", "LB IN-TO", "Left roadbed — Inside Shoulder, Total width"),
    ("Left Roadbed", "LB IN-TR", "Left roadbed — Inside Shoulder, Treated width"),
    ("Median", "Med Eff", "Median — section effective date"),
    ("Median", "Med T", "Median — Type"),
    ("Median", "Med C", "Median — Curb & Landscape"),
    ("Median", "Med B", "Median — Barrier"),
    ("Median", "Med V/WDA", "Median — Width + Variance code (e.g. '14Z', '08V')"),
    ("Right Roadbed", "RB Eff", "Right roadbed — section effective date"),
    ("Right Roadbed", "RB S/T", "Right roadbed — Surface Type"),
    ("Right Roadbed", "RB #Ln", "Right roadbed — Number of Lanes"),
    ("Right Roadbed", "RB S/F", "Right roadbed — Special Feature ('Z' = none)"),
    ("Right Roadbed", "RB IN-TO", "Right roadbed — Inside Shoulder, Total width"),
    ("Right Roadbed", "RB IN-TR", "Right roadbed — Inside Shoulder, Treated width"),
    ("Right Roadbed", "RB Wid", "Right roadbed — Traveled Way Width"),
    ("Right Roadbed", "RB OT-TO", "Right roadbed — Outside Shoulder, Total width"),
    ("Right Roadbed", "RB OT-TR", "Right roadbed — Outside Shoulder, Treated width"),
]

HEADER = [c[1] for c in COLUMNS]            # the 34 export labels (correct as-is)
ROUTE_COL = "Route"                         # leading column on consolidated workbooks

assert len(HEADER) == 34                    # layout guard (the export's 34 columns)


def recognize(header):
    """Is `header` (a loaded row-1 list, possibly with a leading 'Route') the
    Highway Detail layout? Returns has_route (bool) when recognized, else None.
    Comparison is by the full label list, so a same-width but different report
    can't sneak through."""
    if header == list(HEADER):
        return False
    if header == [ROUTE_COL] + list(HEADER):
        return True
    return None


def legend_rows():
    """Rows for the 'Legend' sheet: (group, label, meaning), in column order."""
    return [(grp, label, meaning) for grp, label, meaning in COLUMNS]


def tooltip_for(label):
    """The hover-tooltip text for a header `label` (group prefix + meaning), or
    None when the label isn't a Highway Detail column."""
    for grp, lab, meaning in COLUMNS:
        if lab == label:
            return f"[{grp}] {meaning}" if grp else meaning
    return None


# ---------------------------------------------------------------------------
# openpyxl helpers (guarded — importing this module must never need openpyxl).
# Same shapes as highway_log_columns so the consolidators/comparisons plug in
# identically. Work in both normal and write_only (streaming) workbooks.
# ---------------------------------------------------------------------------
try:
    from openpyxl.comments import Comment
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPX = True
except ImportError:
    _OPX = False

_LEGEND_DARK = "1F3864"
_LEGEND_TITLE = ("Highway Detail — column legend. Each record is two printed "
                 "lines: line 1 the location/record/access/city attributes, "
                 "line 2 the description + the Left Roadbed / Median / Right "
                 "Roadbed blocks, each led by its own effective date.")


def comment_for(label):
    """An openpyxl Comment (hover tooltip) for header `label`, or None."""
    if not _OPX:
        return None
    text = tooltip_for(label)
    if text is None:
        return None
    c = Comment(text, "TSMIS Exporter")
    c.width, c.height = 320, 120
    return c


def write_legend_sheet(wb, title="Legend"):
    """Append a Legend worksheet (Group / Column / Meaning) to `wb`.
    Streaming-safe: only create_sheet + append are used, so it works on a
    write_only workbook too. No-op if openpyxl is unavailable."""
    if not _OPX:
        return None
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = _LEGEND_DARK
    hfont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color=_LEGEND_DARK)
    note_font = Font(name="Arial", size=10, italic=True, color="595959")
    body = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    write_only = getattr(wb, "write_only", False)

    def cell(value, font=body, fill=None, align=None):
        if write_only:
            c = WriteOnlyCell(ws, value=value)
            c.font = font
            if fill:
                c.fill = fill
            if align:
                c.alignment = align
            return c
        return value

    for col, w in (("A", 16), ("B", 14), ("C", 84)):
        ws.column_dimensions[col].width = w
    ws.append([cell(_LEGEND_TITLE, note_font)])
    ws.append([cell(h, hfont, hfill) for h in ("Group", "Column", "Meaning")])
    for grp, label, meaning in legend_rows():
        ws.append([cell(grp or "—"), cell(label), cell(meaning, body, align=wrap)])
    if not write_only:
        for row in ws.iter_rows(min_row=2, max_row=2):
            for c in row:
                c.font = hfont
                c.fill = hfill
    return ws


def apply_header_tooltips(ws, first_row=1):
    """Attach the hover tooltip to each Highway Detail header cell on an
    already-written, NON-streaming sheet. No-op if openpyxl is unavailable."""
    if not _OPX:
        return
    for c in next(ws.iter_rows(min_row=first_row, max_row=first_row)):
        cm = comment_for(c.value)
        if cm is not None:
            c.comment = cm


# ---------------------------------------------------------------------------
# Ditto ('+'-run) resolution — the paired-roadbed convention, Highway Detail's
# own flavor. Same domain rule the Highway Log study established
# (docs/highway_log/comparison-study.md §3): a `+`/`++` cell is a POINTER to the
# paired roadbed's own row, never a copy of the row above, so it is never a
# difference in itself. Only the DISPLAY fill lives here — the non-asserting diff
# is CompareSchema.ditto_nonasserting and needs no value at all.
#
# Censused on the 60,083-row statewide TSN extract (2026-08-18):
#   * 1,992 rows carry a dittoed block: 1,027 dittoed LEFT (all HG='R') and
#     965 dittoed RIGHT (all HG='L') — exactly the study's model, and 100%
#     consistent: zero rows ditto BOTH blocks, zero PARTIAL blocks. The block is
#     always the unit, so the roadbed a row describes is unambiguous.
#   * The paired row is found by SPAN, not by equal postmile: the two roadbeds
#     are segmented independently, so only 27.8% of pairs share a postmile while
#     90.8% are covered by the opposite roadbed's [postmile, postmile+length)
#     span (raw-extract figures).
#   * Measured on the comparison's OWN rows this fills 14,490 of 17,928 cells
#     (80.8%); the rest have no covering span, or two that disagree, and are left
#     UNFILLED rather than guessed (the cell still marks itself as a ditto and
#     says no paired value was found). Lower than the raw figure because the
#     comparison groups by route where the census also split on the PP prefix, so
#     a cross-county route can offer two disagreeing spans — refused on purpose.
#
# Deliberately NOT Highway Log's nearest-row fallback: HD's spans make a
# principled answer available for 9 in 10 cells, and inventing one for the rest
# would trade a visibly-unresolved cell for a quietly-wrong one. The fill is
# informational either way — it can never change a diff result.
# ---------------------------------------------------------------------------
import re as _re
from decimal import Decimal as _Decimal, InvalidOperation as _InvalidOperation

_DITTO_RE = _re.compile(r"^\++$")
# Post Mile / Length / HG positions within the comparison's SHARED_HEADER, and
# the two roadbed blocks' column spans. Kept here beside the labels so the
# comparison module has one place to change if the header ever moves.
_PM_I, _LEN_I, _HG_I = 0, 2, 4
_LB_COLS = tuple(range(12, 21))          # LB Eff .. LB IN-TR
_RB_COLS = tuple(range(26, 35))          # RB Eff .. RB OT-TR


def is_ditto(v):
    """True for a Highway Detail ditto marker — a non-empty run of only '+'."""
    return bool(v is not None and _DITTO_RE.fullmatch(str(v).strip()))


# Mirrors compare_highway_detail_tsn._PM_RE: a postmile is optional PREFIX
# letters, the mile, then optional trailing letters (roadbed R/L, equation E).
# Both ends must come off before the mile can be read as a number — stripping
# only the roadbed suffix leaves 'R081.505', which is not a Decimal.
_PM_PARTS_RE = _re.compile(r"^([A-Z]*?)(\d{1,3}\.\d{1,3})([A-Z]*)$")


def _pm_split(pm):
    """(mile, roadbed) for a canonical Post Mile — ('081.505', 'R') — or
    (None, '') when the token isn't a postmile."""
    m = _PM_PARTS_RE.match(str(pm or "").strip().upper())
    if not m:
        return None, ""
    bed = ""
    for ch in m.group(3):
        if ch in ("R", "L"):
            bed = ch
    return m.group(2), bed


def _roadbed(pm):
    """The roadbed a canonical Post Mile names — its trailing 'R'/'L', or ''."""
    return _pm_split(pm)[1]


def _dec(v):
    try:
        return _Decimal(str(v).strip() or "0")
    except (_InvalidOperation, AttributeError, ValueError):  # silent-ok: pure numeric predicate — a non-numeric postmile/length simply has no span, which the caller skips
        return None


def _mile(pm):
    """The numeric mile of a canonical Post Mile, or None."""
    return _dec(_pm_split(pm)[0]) if _pm_split(pm)[0] is not None else None


def paired_roadbed_fills(rows, has_route):
    """`{row_index: {col_in_row: resolved_or_None}}` for every dittoed roadbed
    cell — the CompareSchema.ditto_resolver contract.

    A row whose LEFT block is dittoed describes the RIGHT roadbed, so its Left
    geometry is the one printed on the LEFT-roadbed row covering this postmile
    (and vice versa). Resolution is per route, by covering span; a cell with no
    covering span, or with spans that disagree, resolves to None and is reported
    as "no paired value found" rather than filled with a guess."""
    off = 1 if has_route else 0
    groups = {}
    for gi, r in enumerate(rows):
        groups.setdefault(r[0] if has_route else "", []).append(gi)

    out = {}
    for members in groups.values():
        # Index the route's rows by which roadbed they describe, with their span.
        spans = {"R": [], "L": []}
        for gi in members:
            r = rows[gi]
            pm = _mile(r[off + _PM_I])
            ln = _dec(r[off + _LEN_I])
            if pm is None or ln is None:
                continue
            bed = _roadbed(r[off + _PM_I]) or str(r[off + _HG_I] or "").strip().upper()
            if bed in ("R", "L"):
                spans[bed].append((pm, pm + ln, gi))

        for gi in members:
            r = rows[gi]
            for cols, want in ((_LB_COLS, "L"), (_RB_COLS, "R")):
                have = [c for c in cols if off + c < len(r)]
                if not have or not all(is_ditto(r[off + c]) for c in have):
                    continue          # only a COMPLETE dittoed block resolves
                pm = _mile(r[off + _PM_I])
                found = set()
                for lo, hi, mate in spans[want]:
                    if mate == gi or pm is None or not (lo <= pm < hi):
                        continue
                    vals = tuple(rows[mate][off + c] if off + c < len(rows[mate])
                                 else None for c in have)
                    if not all(is_ditto(v) for v in vals):
                        found.add(vals)
                one = found.pop() if len(found) == 1 else None
                for n, c in enumerate(have):
                    out.setdefault(gi, {})[c] = one[n] if one is not None else None
    return out
