"""Build the TSMIS-vs-TSN Highway Detail discrepancy workbook (FLAT, route+PM).

The Intersection Detail FLAT recipe applied to Highway Detail: both sides are
XLSX in different shapes, each with its own loader projecting onto ONE shared
PM-keyed header. Reconciled against the full statewide bundle (2026-07: 252
TSMIS routes / 51,243 rows vs the 60,083-row TSN extract; 46,847 rows matched):

  * TSMIS side — the CONSOLIDATED Highway Detail workbook (leading Route column
    + the 34 export columns, labels correct as-is), read BY POSITION.
  * TSN side — the statewide raw `Sheet 1` (56 named DB columns; route from
    RTE + RTE_SFX), or the normalized TSN-library workbook.

Row identity — the CANONICAL Post Mile key. The two systems encode the same
row differently, so the key is rebuilt on both sides:
  * TSMIS glues prefix + mile + a trailing marker into one token ('R012.243',
    'S000.000', '000.000E', '000.080R'); TSN keeps PP / POSTMILE / E_IND as
    separate columns and marks an independent-alignment roadbed via HG (R/L)
    with a BARE postmile.
  * The canonical key = prefix + zero-padded mile + roadbed, where the roadbed
    letter comes from the TSMIS trailing R/L or (both sides) HG∈{R,L}. The
    equation marker 'E' is NOT part of the key — the two systems disagree on
    where they print it (TSMIS 'C043.925R' vs TSN 'C 043.925 E' is the SAME
    row) — it is surfaced as the compared 'PS' column instead, so a marker
    difference flags as a field diff rather than splitting the row one-sided.

Every field present in both systems is COMPARED and COUNTED — a mechanical diff;
the reader adds commentary, the tool never hides a column (the Intersection
Detail position-aligned policy, user decision 2026-06-24). Normalizations make
some raw-different values compare equal; each is documented in the Notes sheet:
  1. **Non-Add ('NA')** — TSN prints an explicit 'A' for ordinary add mileage
     where TSMIS leaves the cell blank (98.7% of matched rows). 'A' normalizes
     to blank so only a genuine flag change (N vs blank) flags.
  2. **Zero-padding** — TSMIS zero-pads single digits ('02' lanes, '08'
     shoulders); TSN doesn't ('2', '8'). Numeric columns compare as numbers.
  3. **Length** — TSN stores raw DB precision (0.01098); TSMIS prints 3
     decimals ('000.011'). Both normalize to the printed 3-decimal form.
  4. **Med V/WDA** — TSN stores Width and Variance separately (14 + 'Z');
     TSMIS glues them ('14Z'). The TSN pair is glued the same way.
  5. **RU Eff** — TSMIS prints the Rural/Urban layer's effective date in the
     slot where the legacy TASAS report prints the ADT profile BEGIN date
     (TSN's BEG_DATE, a Jan-1 count year). Same printed position, different
     meaning — compared by position (nothing suppressed) and it differs on
     ~99% of rows; the Notes sheet says why.
The TSN-only ADT INFORMATION block (LK-AHD / P / LK-BACK / CHANGE/MILE / DVM;
TSMIS omits it by design) and the TSN */Y change flags are not compared; the
ADT block appears in blue on the Report View for reference. A second "Report
View" sheet replicates the printed two-line TASAS record and shows every
difference in red (date-column diffs included, but kept out of its per-record
"Major" count, matching the Intersection Detail replica).

Console-free; engine in compare_core.
"""
import dataclasses
import logging
import re

from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import compare_tsn_common as ctc
from compare_tsn_common import (load_consolidated_rows, row_has_data,
                                suggest_route_name)
from compare_core import (CompareSchema, compared_cell, normalize_value, keys_for,
                          pair_occurrences_by_similarity, union_keys,
                          set_safe_literal_cell, _PROGRESS_EVERY)
import highway_detail_columns as hdc

log = logging.getLogger("tsmis.compare")
REPORT_NAME = "Highway Detail"
TSMIS_SHEET = "Highway Detail"            # consolidated sheet (Route prepended)
TSN_SHEET = "Sheet 1"                     # raw statewide DB dump
NORMALIZED_SHEET = "Highway Detail (TSN)"
# CMP-AUD-034: the EXACT consolidated-TSMIS header (['Route'] + the export's own
# 34-column roadbed layout). _tsmis_row reads by position, so this is bound
# exactly — a shifted/relabelled/wrong-edition header is refused, never mis-mapped.
# Verified statewide-stable (252/252 routes → 1 header) + data-source/edition-
# independent (2026-07-17 census); the Highway Detail (PDF) consolidator emits the
# IDENTICAL 34-column header, so the PDF-vs-Excel self-check that loads side A
# through this same _load_tsmis stays valid.
_TSMIS_HEADER = ["Route", "Post Mile", "Length", "Date of Rec", "HG", "AC",
                 "Acc-Cont Eff", "City", "RU", "RU Eff", "Description", "NA",
                 "LB Eff", "LB S/T", "LB #Ln", "LB S/F", "LB OT-TO", "LB OT-TR",
                 "LB Wid", "LB IN-TO", "LB IN-TR", "Med Eff", "Med T", "Med C",
                 "Med B", "Med V/WDA", "RB Eff", "RB S/T", "RB #Ln", "RB S/F",
                 "RB IN-TO", "RB IN-TR", "RB Wid", "RB OT-TO", "RB OT-TR"]
# CMP-AUD-037: the DIRECT-path freshness marker version (the catalog's
# highway_detail normalization_version MIRRORS this; tsn_load_highway_detail
# .build_into stamps it, _load_tsn refuses anything older). HD's direct loader
# had NO stale-library gate at all — a normalized workbook from any prior
# normalizer was trusted. v3 is a marker-only bump over the v2 sidecar shape;
# the rows are byte-identical, and the bump forces D2 to rebuild stored
# libraries so they gain the marker.
NORMALIZATION_VERSION = 3
# CMP-AUD-033: the documented sidecar columns that follow ["Route"] +
# SHARED_HEADER in the normalized workbook (tsn_load_highway_detail.SIDECAR_HEADER
# mirrors this; check_tsn_normalization_marker gates the mirror).
_NORMALIZED_SIDECARS = ("TSN District", "TSN County")

KEY = "Post Mile"
# The shared comparison header: the canonical Post Mile key, the derived PS
# (equation-marker) column, then the remaining 33 TSMIS export columns in
# report order. Nothing is suppressed (CONTEXT_FIELDS is empty).
SHARED_HEADER = [
    "Post Mile", "PS", "Length", "Date of Rec", "HG", "AC", "Acc-Cont Eff",
    "City", "RU", "RU Eff", "Description", "NA",
    "LB Eff", "LB S/T", "LB #Ln", "LB S/F", "LB OT-TO", "LB OT-TR", "LB Wid",
    "LB IN-TO", "LB IN-TR",
    "Med Eff", "Med T", "Med C", "Med B", "Med V/WDA",
    "RB Eff", "RB S/T", "RB #Ln", "RB S/F", "RB IN-TO", "RB IN-TR", "RB Wid",
    "RB OT-TO", "RB OT-TR",
]
KEY_FIELD = SHARED_HEADER.index(KEY)      # 0
CONTEXT_FIELDS = ()                       # position-aligned: nothing suppressed
DATE_FIELDS = ("Date of Rec", "Acc-Cont Eff", "RU Eff", "LB Eff", "Med Eff",
               "RB Eff")
# Zero-padded numeric columns (TSMIS '02' vs TSN '2') — compared as numbers.
NUMERIC_FIELDS = ("LB #Ln", "LB OT-TO", "LB OT-TR", "LB Wid", "LB IN-TO",
                  "LB IN-TR", "RB #Ln", "RB OT-TO", "RB OT-TR", "RB Wid",
                  "RB IN-TO", "RB IN-TR")

# TSN raw column name for each DIRECTLY-mapped shared field (the multi-column
# derivations — Post Mile, PS, Med V/WDA and the route — are built in _tsn_row).
_TSN_COL = {
    "Length": "LENGTH", "Date of Rec": "REC_DATE", "HG": "HG", "AC": "AC",
    "Acc-Cont Eff": "ACC_EFF_DATE", "City": "CITY", "RU": "POP_CODE",
    "RU Eff": "BEG_DATE", "Description": "DESCRIPTION", "NA": "NON_ADD",
    "LB Eff": "L_EFF_DATE", "LB S/T": "L_ST", "LB #Ln": "L_NO_LANES",
    "LB S/F": "L_SF", "LB OT-TO": "L_OT_TOT", "LB OT-TR": "L_OT_TR",
    "LB Wid": "L_TR_WID", "LB IN-TO": "L_IN_TOT", "LB IN-TR": "L_IN_TR",
    "Med Eff": "M_EFF_DATE", "Med T": "M_TYPE_CODE", "Med C": "M_CL",
    "Med B": "M_BA",
    "RB Eff": "R_EFF_DATE", "RB S/T": "R_ST", "RB #Ln": "R_NO_LANES",
    "RB S/F": "R_SF", "RB IN-TO": "R_IN_TOT", "RB IN-TR": "R_IN_TR",
    "RB Wid": "R_TR_WID", "RB OT-TO": "R_OT_TOT", "RB OT-TR": "R_OT_TR",
}
# Consolidated-TSMIS VALUE position for each export column (Route at 0, then
# the 34 export columns in order — the header labels are correct, but position
# is authoritative like the other consolidated loaders).
_TSMIS_POS = {name: i + 1 for i, name in enumerate(hdc.HEADER)}
_TSMIS_HG_POS = _TSMIS_POS["HG"]          # the roadbed fallback reads HG


# --------------------------------------------------------------------------- #
# normalization
# --------------------------------------------------------------------------- #
def _s(v):
    """None-safe stripped text (a numeric 0 stays '0', never blank)."""
    return "" if v is None else str(v).strip()


_PM_RE = re.compile(r"^([A-Z]*?)(\d{1,3}\.\d{1,3})([A-Z]*)$")


def _pm_parts(token):
    """Split a glued postmile token into (prefix, mile, trailing letters);
    an unrecognized token comes back as ('', token, '')."""
    m = _PM_RE.match(_s(token).upper())
    if not m:
        return "", _s(token), ""
    return m.group(1), m.group(2), m.group(3)


def _pad_mile(mile):
    """'0.08' / '11.228' -> the report's fixed '000.080' / '011.228' form."""
    try:
        return f"{float(mile):07.3f}"
    except (TypeError, ValueError):
        return _s(mile)


def pm_canon(token, hg):
    """The canonical roadbed-aware Post Mile KEY: <prefix><MMM.mmm><roadbed>.

    The roadbed letter comes from the token's own trailing R/L (TSMIS glues it)
    or, when the token has none, from HG∈{R,L} (how TSN marks an independent-
    alignment row; also the TSMIS fallback when the single trailing-marker slot
    is taken by 'E'). The equation marker is NOT part of the key — the systems
    disagree on where they print it (see the module docstring); it is compared
    as the separate PS column instead."""
    prefix, mile, trail = _pm_parts(token)
    hg = _s(hg).upper()
    roadbed = ""
    for ch in trail:
        if ch in ("R", "L"):
            roadbed = ch
    if not roadbed and hg in ("R", "L"):
        roadbed = hg
    return f"{prefix}{_pad_mile(mile)}{roadbed}"


def pm_suffix(token, e_ind=None):
    """The PS (equation-marker) column value: an explicit E_IND (TSN) or an 'E'
    inside the glued token's trailing letters (TSMIS / the canonical form)."""
    if _s(e_ind).upper() == "E":
        return "E"
    _p, _m, trail = _pm_parts(token)
    return "E" if "E" in trail else ""


def _norm_len(v):
    """Length to the printed 3-decimal form ('000.011'): TSN stores raw DB
    precision (0.01098) where TSMIS prints the fixed 3-decimal mile."""
    s = _s(v)
    if not s:
        return ""
    try:
        return f"{float(s):07.3f}"
    except ValueError:
        return s


def _norm_date(v):
    """Both systems print YY-MM-DD text; a datetime-typed cell (openpyxl) is
    rendered to the same form so the type can't fake a difference."""
    if hasattr(v, "strftime"):
        return v.strftime("%y-%m-%d")
    return _s(v)


def _norm_num(v):
    """Zero-padded count/width to a canonical number: TSMIS '02' = TSN '2',
    '00' = '0'. A real numeric 0 stays '0', never blank (the falsy-zero trap);
    non-numeric values pass through unchanged."""
    s = _s(v)
    if s.isdigit():
        return str(int(s))
    return s


def _norm_na(v):
    """Non-Add: TSN prints an explicit 'A' for ordinary ADD mileage where the
    TSMIS report leaves the cell blank — fold 'A' to blank (both sides; TSMIS
    never prints an 'A') so only a genuine N-vs-blank change flags."""
    s = _s(v).upper()
    return "" if s == "A" else s


def _norm_wda(v):
    """Med V/WDA to the TSMIS glued form: 2-digit-padded width + variance
    letter ('8V' -> '08V', '14Z' stays). Non-matching shapes pass through."""
    s = _s(v).upper()
    m = re.fullmatch(r"(\d+)([A-Z]?)", s)
    if not m:
        return s
    return f"{int(m.group(1)):02d}{m.group(2)}"


_WS_RE = re.compile(r"\s+")


def _norm_desc(v):
    """Descriptions: collapse whitespace runs (the TSN extract pads with fixed-
    width blanks; TSMIS joins multiple landmarks with ', ')."""
    return _WS_RE.sub(" ", _s(v))


def _v(x):
    return normalize_value(x)


def _project(field, raw):
    """Normalize one raw cell for `field` into the shared, comparable form.
    (Post Mile / PS are multi-column derivations handled by the row loaders.)"""
    if field in DATE_FIELDS:
        return _norm_date(raw)
    if field in NUMERIC_FIELDS:
        return _norm_num(raw)
    if field == "Length":
        return _norm_len(raw)
    if field == "NA":
        return _norm_na(raw)
    if field == "Med V/WDA":
        return _norm_wda(raw)
    if field == "Description":
        return _norm_desc(raw)
    return _s(_v(raw))


def _norm_route_token(rte, sfx=""):
    """Route token matching the TSMIS per-route filenames: 3-digit-padded
    number + upper-cased suffix ('1'->'001', '5','S'->'005S')."""
    s = _s(rte)
    m = re.fullmatch(r"(\d+)([A-Za-z]?)", s)
    base, tail = (m.group(1), m.group(2)) if m else (s, "")
    tail = (tail or "") + _s(sfx)
    if base.isdigit():
        base = f"{int(base):03d}"
    return f"{base}{tail}".upper()


TSN_RAW_HEADER = (
    "THY_ID", "DIST", "CNTY", "RTE", "RTE_SFX", "DIST_CNTY_ROUTE", "PP",
    "POSTMILE", "E_IND", "LENGTH", "REC_DATE", "HG", "AC", "ACC_SIG",
    "ACC_EFF_DATE", "CITY", "POP_CODE", "BEG_DATE", "ADT_AMT", "PROFILE",
    "BREAK_DESC", "LK_BACK_ADT", "CHNGMILE", "DVM", "DESCRIPTION", "NON_ADD",
    "LT_SIG", "L_EFF_DATE", "L_ST", "L_NO_LANES", "L_SF", "L_OT_TOT",
    "L_OT_TR", "L_TR_WID", "L_IN_TOT", "L_IN_TR", "MED_SIG", "M_EFF_DATE",
    "M_TYPE_CODE", "M_CL", "M_BA", "M_WID", "M_VA", "RT_SIG", "R_EFF_DATE",
    "R_ST", "R_NO_LANES", "R_SF", "R_IN_TOT", "R_IN_TR", "R_TR_WID",
    "R_OT_TOT", "R_OT_TR", "SEG_ORDER_ID", "REFERENCE_DATE", "EXTRACT_DATE",
)


# --------------------------------------------------------------------------- #
# loaders -> consolidated-shape rows ([route, *SHARED_HEADER])
# --------------------------------------------------------------------------- #
def _tsn_row(r, h):
    def g(name):
        i = h.get(name)
        return r[i] if i is not None and i < len(r) else None
    route = _norm_route_token(g("RTE"), g("RTE_SFX"))
    token = f"{_s(g('PP'))}{_s(g('POSTMILE'))}"
    row = [route]
    for f in SHARED_HEADER:
        if f == "Post Mile":
            row.append(pm_canon(token, g("HG")))
        elif f == "PS":
            row.append(pm_suffix(token, g("E_IND")))
        elif f == "Med V/WDA":
            row.append(_norm_wda(f"{_s(g('M_WID'))}{_s(g('M_VA'))}"))
        else:
            row.append(_project(f, g(_TSN_COL[f])))
    return row


def require_tsn_raw_header(header):
    ctc.require_exact_raw_header(header, TSN_RAW_HEADER, REPORT_NAME)


def tsn_rows_from_raw(path):
    with ctc.exact_raw_rows(
            path, TSN_SHEET, TSN_RAW_HEADER, REPORT_NAME,
            required_nonblank=("DIST", "CNTY", "RTE", "POSTMILE")) as (header, rows_in):
        h = {n: i for i, n in enumerate(header)}
        return [_tsn_row(r, h) for r in rows_in]


def _normalized_row(r):
    """Re-project one row from the normalized TSN-library sheet onto the shared
    shape, RE-APPLYING the field normalizations. The projections are idempotent
    on already-normalized values, so this is a no-op for a freshly-built library
    BUT repairs a STALE one (a library built before a normalization change would
    otherwise feed old values straight through) — normalization changes take
    effect immediately, without waiting for the D2 rebuild."""
    vals = list(r)[:len(SHARED_HEADER) + 1]
    vals += [None] * (len(SHARED_HEADER) + 1 - len(vals))
    hg = _s(vals[1 + SHARED_HEADER.index("HG")])
    out = [_s(_v(vals[0]))]
    for i, f in enumerate(SHARED_HEADER):
        raw = vals[i + 1]
        if f == "Post Mile":
            out.append(pm_canon(raw, hg))
        elif f == "PS":
            out.append(pm_suffix(raw))
        else:
            out.append(_project(f, raw))
    return out


def _load_tsn(path):
    path = Path(path)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {path.name}: {type(e).__name__}: {e}")
    try:
        if NORMALIZED_SHEET in wb.sheetnames:
            it = wb[NORMALIZED_SHEET].iter_rows(values_only=True)
            header = [("" if c is None else str(c).strip())
                      for c in (next(it, None) or ())]
            # CMP-AUD-033: bind the header to the exact ["Route"] + SHARED_HEADER
            # prefix + documented sidecars before reading BY POSITION — HD's
            # loader trusted any normalized sheet, so a reordered/renamed header
            # silently mis-mapped every column.
            ctc.require_shared_header_prefix(
                header, ["Route"] + SHARED_HEADER, _NORMALIZED_SIDECARS,
                path.name, REPORT_NAME)
            # CMP-AUD-037: HD's direct loader had no freshness gate — a normalized
            # workbook from any prior normalizer was trusted. Refuse a pre-v3
            # library (no in-workbook marker); the library path auto-rebuilds (D2).
            ctc.require_current_normalization(
                wb, path.name, NORMALIZATION_VERSION,
                "pre-v3: no in-workbook normalization marker")
            rows = [_normalized_row(r)
                    for r in it if r and any(c not in (None, "") for c in r)]
            return rows, True
    finally:
        wb.close()
    return tsn_rows_from_raw(path), True


def _tsmis_row_with(r, project, extra=None):
    """One consolidated TSMIS row with `project(field, raw)` supplying the
    value projection and `extra(at, token)` optionally appending trailing
    cells. The canonical roadbed-aware Post Mile KEY and the PS derivation
    are IDENTICAL for every caller (pairing semantics are shared); only the
    value projection varies — `_project` for the vs-TSN comparison
    (cross-system reconciliation), a verbatim projection for the same-source
    PDF-vs-Excel flavor, which also appends the RAW printed Post Mile as its
    own compared cell so a dropped roadbed letter can no longer hide inside
    the canonical key (CMP-AUD-067)."""
    def at(i):
        return r[i] if i < len(r) else None
    token = _s(at(_TSMIS_POS["Post Mile"]))
    hg = at(_TSMIS_HG_POS)
    row = [_s(_v(at(0)))]
    for f in SHARED_HEADER:
        if f == "Post Mile":
            row.append(pm_canon(token, hg))
        elif f == "PS":
            row.append(pm_suffix(token))
        else:
            row.append(project(f, at(_TSMIS_POS[f])))
    if extra is not None:
        row.extend(extra(at, token))
    return row


def _tsmis_row(r):
    return _tsmis_row_with(r, _project)


def _load_tsmis(path):
    return load_consolidated_rows(
        path, TSMIS_SHEET,
        missing_sheet_hint="pick the consolidated TSMIS Highway Detail workbook.",
        bad_header_msg="isn't a CONSOLIDATED Highway Detail workbook in the current "
                       "site layout (expected a leading 'Route' column and the exact "
                       "34-column export header) — consolidate a fresh export first.",
        header_ok=ctc.exact_consolidated_header_ok(_TSMIS_HEADER),  # CMP-AUD-034
        row_transform=_tsmis_row)


# --------------------------------------------------------------------------- #
# Notes sheet — documents every normalization applied (so a match is read as
# "equal after the stated normalization", not raw equality) and comments on the
# columns that differ wholesale. Nothing is suppressed.
# --------------------------------------------------------------------------- #
def _write_notes_sheet(wb):
    ws = wb.create_sheet("Notes")
    ws.sheet_properties.tabColor = "ED7D31"
    write_only = getattr(wb, "write_only", False)
    title = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    head = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color="1F3864")
    sec_fill = PatternFill("solid", start_color="0070C0")
    body = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)

    def cell(value, font=body, f=None, align=None):
        if not write_only:
            return value
        c = WriteOnlyCell(ws, value=value)
        c.font = font
        if f:
            c.fill = f
        if align:
            c.alignment = align
        return c

    def section(text):
        ws.append([cell(text, head, f=sec_fill)])

    def note(text):
        ws.append([cell(text, body, align=wrap)])

    ws.column_dimensions["A"].width = 110
    ws.append([cell("Highway Detail — TSMIS vs TSN: comparison notes", title, fill)])
    note("This is a MECHANICAL field-by-field diff. EVERY column present in both systems "
         "is compared and counted — nothing is hidden. Some columns differ on nearly every "
         "row by their nature; they are still flagged, and the reasons are noted below.")

    section("ROW IDENTITY  (how the same physical row pairs across the two encodings)")
    note("• Rows are keyed on Route + a CANONICAL Post Mile: prefix + zero-padded mile + "
         "roadbed letter. TSMIS glues an R/L onto the postmile for an independent-alignment "
         "roadbed row ('000.080R'); TSN prints the bare postmile and says R/L in the HG "
         "column instead — the canonical key unifies the two (HG also supplies the roadbed "
         "when TSMIS's single trailing-marker slot is occupied by 'E').")
    note("• The equation marker 'E' is NOT part of the key: the two systems disagree on "
         "where they print it (TSMIS 'C043.925R' and TSN 'C 043.925 E' are the SAME row). "
         "It is compared in the separate 'PS' column, so a marker difference flags there "
         "instead of splitting the row into a false one-sided pair.")

    section("NORMALIZATIONS APPLIED  (these can make raw-different values compare EQUAL — "
            "a match means 'equal after the normalization named here', not that the source "
            "cells were byte-identical)")
    note("• Non-Add ('NA') — TSN prints an explicit 'A' on ordinary add-mileage rows where "
         "the TSMIS report leaves the cell blank (98.7% of matched rows statewide). 'A' is "
         "folded to blank, so only a genuine flag change (N vs blank) counts. HOW TO SEE "
         "IT: a blank NA on the TSN side may have been stored 'A'.")
    note("• Zero-padding — TSMIS zero-pads single digits in the lane/shoulder/width columns "
         "('02' lanes, '08' shoulders); TSN doesn't ('2', '8'). These columns compare as "
         "numbers, so only a real value change flags.")
    note("• Length — TSN stores raw database precision (e.g. 0.01098 mi); TSMIS prints the "
         "fixed 3-decimal mile ('000.011'). Both sides normalize to the printed 3-decimal "
         "form. Note the Length is 'distance to the NEXT record', so wherever one system "
         "splits a segment the other doesn't, the lengths around the split genuinely differ.")
    note("• Med V/WDA — TSN stores the median Width and Variance separately (14 + 'Z'); "
         "TSMIS glues them into one code ('14Z'). The TSN pair is glued the same way, and "
         "the comparison ignores zero-padding inside the code ('0Z' = '00Z', '6V' = '06V').")
    note("• Dates — both systems print YY-MM-DD; values compare as printed (a date-typed "
         "cell is rendered to the same text first).")
    note("• Descriptions — whitespace runs collapse (the TSN extract pads with fixed-width "
         "blanks). TSMIS often concatenates MULTIPLE landmark descriptions with ', ' where "
         "TSN prints only the first — those still flag, and are a render difference rather "
         "than a data conflict.")

    section("COLUMNS THAT DIFFER WHOLESALE  (compared and counted like any other — the "
            "difference is structural, explained here, NOT a per-row data error)")
    note("• RU Eff — TSMIS prints the Rural/Urban (Population Code) layer's effective date "
         "in the slot where the legacy TASAS report prints the ADT profile BEGIN date (a "
         "Jan-1 count year, TSN's BEG_DATE). Same printed position, different meaning — the "
         "column differs on ~99% of matched rows. Compared by position (nothing is "
         "suppressed); read the count as 'the slot means different things', not as "
         "thousands of data errors.")
    note("• Section effective dates (LB / Med / RB Eff, Acc-Cont Eff) — the two systems "
         "record their OWN effective date for the section carrying the current attributes, "
         "and they disagree on roughly a third of rows. A date difference alone usually "
         "means the systems refreshed the section at different times, not that the "
         "attributes conflict (the attribute columns say that).")
    note("• Date of Rec — mostly matches; TSMIS leaves it blank on its attribute-change / "
         "county-boundary marker rows (~700 statewide), which then flag as blank-vs-value.")
    note("• Special Feature (LB/RB S/F) — 'Z' means 'no special feature'; TSMIS renders a "
         "blank where TSN prints the explicit Z on ~5% of rows. Those blank-vs-Z cells are "
         "counted (a completeness gap, not a geometry conflict). A handful of TSN cells "
         "carry a '+' continuation mark from the source report; they flag visibly.")
    note("• Median T/C/B — TSMIS renders the three median codes from one packed value and "
         "collapses a missing slot (a missing Type shifts Curb/Barrier left), so a row "
         "missing one median code can misalign the other two against TSN's clean columns.")

    section("NOT COMPARED  (present in only one system)")
    note("• The ADT INFORMATION block (LK-AHD / P / LK-BACK / CHANGE-MILE / DVM) — the "
         "TSMIS Highway Detail omits it by design ('the legacy ADT Information block is "
         "omitted'). Shown in blue on the Report View for reference.")
    note("• TSN's change flags (the '*' / 'Y' prefixes on effective dates in the printed "
         "TSN report; separate *_SIG columns in the extract) and its database ids "
         "(THY_ID / SEG_ORDER_ID / REFERENCE_DATE / EXTRACT_DATE).")
    note("• TSN carries ~20 unconstructed/unsigned routes TSMIS doesn't export at all — "
         "their rows appear as 'entire route' one-sided (blue), not as differences.")

    section("REPORT VIEW  (a second sheet — the printed two-line record, for visual inspection)")
    note("• The 'Report View' tab replicates the printed TASAS Highway Detail record (two "
         "physical lines per row) and renders EVERY difference in red — the date columns "
         "included — so the page can be eyeballed straight against the source reports. Per "
         "record it shows two counts: 'Major' = genuine NON-date attribute conflicts (the "
         "date columns and the PS marker are excluded so they don't drown out the real "
         "conflicts); 'Diffs' = every difference. The TSN-only ADT block and the TSN "
         "district-county-route appear there in blue for reference.")
    note("Rows are keyed on Route + the canonical Post Mile.")
    return ws


_SCHEMA = CompareSchema(
    report_name=REPORT_NAME,
    header=SHARED_HEADER,
    side_a="TSMIS",
    side_b="TSN",
    id_noun="location",
    id_noun_plural="locations",
    pair_noun="postmile",
    sides_noun="systems",
    medwid_fields=("Med V/WDA",),
    date_fields=DATE_FIELDS,
    data_widths={"Post Mile": 12, "Description": 26, "Date of Rec": 11},
    cmp_widths={"Post Mile": 12, "Description": 30, "Date of Rec": 12},
    one_sided_note_extra=" (TSN segment splits and ADT profile breaks, TSMIS "
                         "realignment markers, and the unconstructed routes "
                         "TSMIS doesn't export)",
    key_field=KEY_FIELD,
    context_fields=CONTEXT_FIELDS,
    legend_writer=_write_notes_sheet,
    source_file_a=("highway_detail", TSMIS_SHEET, "xlsx"),   # Source Files sheet
)


# --------------------------------------------------------------------------- #
# Report View — a two-line replica of the printed TASAS record, comparison-
# coloured (the Intersection Detail replica pattern; see that module).
# Line 1 = the location/record/access/city attributes + the TSN-only ADT block
# (blue); line 2 = description + NA + the Left Roadbed / Median / Right Roadbed
# blocks. RED = any difference; the date columns and the PS marker classify
# 'soft' (red but excluded from the per-record Major count); BLUE = TSN-only.
# Identity repeats on both physical rows so a filter keeps records together.
# --------------------------------------------------------------------------- #
_RV_ONE = {"ADT": "ADT_AMT", "PROF": "PROFILE", "LKBK": "LK_BACK_ADT",
           "CHG": "CHNGMILE", "DVM": "DVM", "DCR": None}   # DCR is derived
_RV_SOFT = set(DATE_FIELDS) | {"PS"}
_RV_AUX = ("Major", "Diffs", "Route")
# (g1, l1, spec1, g2, l2, spec2) — line-1 group/label/spec stacked over line-2's.
_RV_GRID = [
    ("", "POST MILE", ("pm", None), "", "DESCRIPTION", ("cmp", "Description")),
    ("", "PS", ("cmp", "PS"), "", "NA", ("cmp", "NA")),
    ("", "LENGTH", ("cmp", "Length"), "* LEFT ROADBED *", "EFF-DATE", ("cmp", "LB Eff")),
    ("", "DATE OF REC", ("cmp", "Date of Rec"), "* LEFT ROADBED *", "S/T", ("cmp", "LB S/T")),
    ("", "H/G", ("cmp", "HG"), "* LEFT ROADBED *", "# LN", ("cmp", "LB #Ln")),
    ("", "A/C", ("cmp", "AC"), "* LEFT ROADBED *", "S/F", ("cmp", "LB S/F")),
    ("", "ACC-CONT EFF", ("cmp", "Acc-Cont Eff"), "* LEFT ROADBED *", "OT-TO", ("cmp", "LB OT-TO")),
    ("", "CITY", ("cmp", "City"), "* LEFT ROADBED *", "OT-TR", ("cmp", "LB OT-TR")),
    ("", "R/U", ("cmp", "RU"), "* LEFT ROADBED *", "T-W WID", ("cmp", "LB Wid")),
    ("", "EFF-DATE", ("cmp", "RU Eff"), "* LEFT ROADBED *", "IN-TO", ("cmp", "LB IN-TO")),
    ("TSN only", "DCR", ("tn", "DCR"), "* LEFT ROADBED *", "IN-TR", ("cmp", "LB IN-TR")),
    ("ADT (TSN only)", "LK-AHD", ("tn", "ADT"), "* MEDIAN *", "EFF-DATE", ("cmp", "Med Eff")),
    ("ADT (TSN only)", "P", ("tn", "PROF"), "* MEDIAN *", "T", ("cmp", "Med T")),
    ("ADT (TSN only)", "LK-BACK", ("tn", "LKBK"), "* MEDIAN *", "C", ("cmp", "Med C")),
    ("ADT (TSN only)", "CHG/MILE", ("tn", "CHG"), "* MEDIAN *", "B", ("cmp", "Med B")),
    ("ADT (TSN only)", "DVM", ("tn", "DVM"), "* MEDIAN *", "V/WDA", ("cmp", "Med V/WDA")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "EFF-DATE", ("cmp", "RB Eff")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "S/T", ("cmp", "RB S/T")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "# LN", ("cmp", "RB #Ln")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "S/F", ("cmp", "RB S/F")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "IN-TO", ("cmp", "RB IN-TO")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "IN-TR", ("cmp", "RB IN-TR")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "T-W WID", ("cmp", "RB Wid")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "OT-TO", ("cmp", "RB OT-TO")),
    ("", "", ("blank", None), "* RIGHT ROADBED *", "OT-TR", ("cmp", "RB OT-TR")),
]
# (normal, ALT) fill hex pairs — whole-record zebra bands; 'soft' shares the
# hard RED palette (every date discrepancy renders red) while staying out of
# the Major count. Identical palette to the Intersection Detail replica.
_RV_FILLS = {"hard": ("F8D4D4", "E8B6B6"), "soft": ("F8D4D4", "E8B6B6"),
             "tn": ("DCE5F3", "B7CCE7"), "tm": ("F8E4CF", "E3C9A2"),
             "id": ("FFFFFF", "CFD6DE"), "count": ("FFFFFF", "CFD6DE"),
             "eq": ("FFFFFF", "CFD6DE")}
_RV_FONTCOL = {"hard": "9C0006", "soft": "9C0006", "tn": "163A63", "tm": "7A431A"}
_RV_COMMENTS = {
    "NA": "NORMALIZED: TSN's explicit 'A' (add mileage) = TSMIS blank; only a "
          "genuine N-vs-blank change flags.",
    "PS": "The equation marker, compared separately: the two systems disagree on "
          "where they print 'E', so it flags here (soft) instead of splitting the "
          "row. Excluded from the Major count.",
    "EFF-DATE": "Date columns render RED when they differ but are kept OUT of the "
                "Major count — the systems often recorded different effective "
                "dates for the same attributes.",
    "DATE OF REC": "Mostly matches; blank on TSMIS's attribute-change/county-"
                   "boundary marker rows. Red when different, not Major.",
    "LENGTH": "Distance to the NEXT record — genuinely differs wherever one "
              "system splits a segment the other doesn't.",
    "V/WDA": "NORMALIZED: TSN's separate Width+Variance are glued to the TSMIS "
             "code form ('14Z'); zero-padding inside the code is ignored.",
    "# LN": "NORMALIZED: zero-padding ignored (TSMIS '02' = TSN '2').",
}


def _rv_classify(field):
    """'soft' = a date column or the PS marker (red but excluded from Major);
    'hard' = a genuine attribute conflict (counts as Major)."""
    return "soft" if field in _RV_SOFT else "hard"


def _tsn_onesided(path):
    """Raw TSN one-sided columns (the ADT block + the district-county-route),
    aligned to the rows `tsn_rows_from_raw` yields. Returns None for a
    normalized-library workbook (those columns aren't stored there) — the
    replica then shows the TSN-only cells blank."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log.info("report view: TSN one-sided read failed (%s: %s)",
                 type(e).__name__, str(e).splitlines()[0] if str(e) else "")
        return None
    try:
        if NORMALIZED_SHEET in wb.sheetnames:
            return None
        sn = TSN_SHEET if TSN_SHEET in wb.sheetnames else wb.sheetnames[0]
        it = wb[sn].iter_rows(values_only=True)
        hdr = next(it, None) or []
        h = {str(n).strip(): i for i, n in enumerate(hdr) if n is not None}
        out = []
        for r in it:
            if not row_has_data(r):
                continue
            def gv(col):
                i = h.get(col)
                return "" if i is None or i >= len(r) or r[i] is None else str(r[i]).strip()
            one = {k: gv(col) for k, col in _RV_ONE.items() if col}
            one["DCR"] = gv("DIST_CNTY_ROUTE") or "-".join(
                t for t in (gv("DIST"), gv("CNTY"), gv("RTE") + gv("RTE_SFX")) if t)
            out.append(one)
        return out
    finally:
        wb.close()


def _write_report_view(wb, ctx, tsn_one):
    """Append the two-line 'Report View' — a faithful replica of the printed
    TASAS Highway Detail record — to the streaming comparison workbook. The
    structure (zebra bands, one-sided side-coloured records, Major/Diffs aux
    columns, merged group headers) mirrors the Intersection Detail replica."""
    sc = ctx["sc"]
    events = ctx.get("events")
    rows_a, rows_b = ctx["rows_a"], ctx["rows_b"]
    ka, kb, union = ctx.get("keys_a"), ctx.get("keys_b"), ctx.get("union")
    if ka is None or kb is None or union is None:
        is_cancelled = events.is_cancelled if events is not None else None
        ka = keys_for(
            rows_a, True, key_field=sc.key_field,
            is_cancelled=is_cancelled)
        kb = keys_for(
            rows_b, True, key_field=sc.key_field,
            is_cancelled=is_cancelled)
        pairing = pair_occurrences_by_similarity(
            sc, rows_a, rows_b, ka, kb, True, events)
        if pairing.pairing_quality != "exact":
            raise ValueError(
                "Report View cannot discard capped duplicate-pairing state")
        ka, kb = pairing.keys_a, pairing.keys_b
        union = union_keys(ka, kb, is_cancelled)
    if events is not None:
        events.on_log(f"  Building the Report View tab ({len(union):,} records)…")
    amap = {k: i for i, k in enumerate(ka)}
    bmap = {k: j for j, k in enumerate(kb)}
    field_index = {name: i for i, name in enumerate(sc.header)}
    fi = {name: 1 + i for i, name in enumerate(sc.header)}   # +1 leading route col
    NA_, NG = len(_RV_AUX), len(_RV_GRID)
    NC = NA_ + NG

    def aval(row, name):
        if row is None:
            return ""
        v = row[fi[name]]
        return "" if v is None else str(v).strip()

    Fn = lambda **k: Font(name="Consolas", **{"size": 8.5, **k})
    fill = lambda c: PatternFill("solid", fgColor=c)
    HEAD, GRP = fill("21344F"), fill("3A5688")
    thin = Side(style="thin", color="D2D2D2")
    med = Side(style="medium", color="51607A")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctrW = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")
    _FONTS = {"hard": dict(color=_RV_FONTCOL["hard"], bold=True),
              "soft": dict(color=_RV_FONTCOL["soft"], bold=True),
              "tn": dict(color=_RV_FONTCOL["tn"]), "tm": dict(color=_RV_FONTCOL["tm"]),
              "id": dict(bold=True), "count": dict(bold=True)}
    _BD_NORM = Border(left=thin, right=thin)
    _BD_BOTTOM = Border(left=thin, right=thin, bottom=med)
    _FILL_CACHE = {(st, a): fill(cols[1 if a else 0])
                   for st, cols in _RV_FILLS.items() for a in (False, True)}
    _FONT_CACHE = {st: Fn(**kw) for st, kw in _FONTS.items()}
    _FONT_DEFAULT = Fn()

    ws = wb.create_sheet("Report View")
    ws.sheet_properties.tabColor = "21344F"
    ws.freeze_panes = "F5"      # aux (Major/Diffs/Route) + POST MILE + PS stay in view

    def value(spec, ra, rb, one):
        kind, ref = spec
        if kind == "tn":
            return (one.get(ref, "") if one else "", "tn")
        if kind == "cmp":
            cell = compared_cell(sc, field_index[ref], ra, rb, off=1)
            if not cell.asserting or cell.equal:
                return (cell.display, "eq")
            tm, tn = cell.display_a, cell.display_b
            return (f"{tm or '·'} ≠ {tn or '·'}", _rv_classify(ref))
        return ("", "blank")

    def woc(val, status, alt, *, bottom=False, align=None):
        c = set_safe_literal_cell(WriteOnlyCell(ws), val)
        c.alignment = align or (lft if status == "id" else ctr)
        c.border = _BD_BOTTOM if bottom else _BD_NORM
        c.fill = _FILL_CACHE.get((status, bool(alt)), _FILL_CACHE[("eq", bool(alt))])
        c.font = _FONT_CACHE.get(status, _FONT_DEFAULT)
        return c

    def hcell(val, fillc, font, align, comment_ref=None):
        c = WriteOnlyCell(ws, value=val)
        c.fill = fillc; c.font = font; c.alignment = align; c.border = bd
        t = _RV_COMMENTS.get(comment_ref) if comment_ref else None
        if t:
            cm = Comment(t, "TSMIS vs TSN"); cm.width = 250; cm.height = 130
            c.comment = cm
        return c

    # ---- 4-row header (the report's two stacked header blocks) ----
    g1 = [col[0] for col in _RV_GRID]
    g2 = [col[3] for col in _RV_GRID]

    def group_row(groups):
        cells, i = [], 0
        while i < NG:
            g, j = groups[i], i
            while j < NG and groups[j] == g:
                j += 1
            for k in range(i, j):
                if g:
                    cells.append(hcell(g if k == i else "", GRP,
                                       Fn(bold=True, color="FFFFFF", size=7.5), ctr))
                else:
                    cells.append(hcell("", PatternFill(), Fn(), ctr))
            i = j
        return cells

    aux_white = Fn(bold=True, color="FFFFFF", size=8)
    blank_dark = Fn()
    ws.append([hcell(lab, HEAD, aux_white, ctrW) for lab in _RV_AUX] + group_row(g1))
    ws.append([hcell("", HEAD, blank_dark, ctrW) for _ in _RV_AUX] +
              [hcell(col[1], HEAD, Fn(bold=True, color="FFFFFF", size=7.5), ctrW, col[1])
               for col in _RV_GRID])
    ws.append([hcell("", HEAD, blank_dark, ctrW) for _ in _RV_AUX] + group_row(g2))
    ws.append([hcell("", HEAD, blank_dark, ctrW) for _ in _RV_AUX] +
              [hcell(col[4], HEAD, Fn(color="BBD0EC", size=7.5), ctrW,
                     col[5][1] if col[5][0] == "cmp" else col[4])
               for col in _RV_GRID])

    # ---- data: two physical rows per record, whole-record alternating band ----
    for n, key in enumerate(union):
        if events is not None and n and n % _PROGRESS_EVERY == 0:
            events.on_log(f"  Report View: {n:,} of {len(union):,} records…")
        ra = rows_a[amap[key]] if key in amap else None
        rb = rows_b[bmap[key]] if key in bmap else None
        one = (tsn_one[bmap[key]] if (tsn_one and key in bmap and bmap[key] < len(tsn_one)) else {})
        pmval = key[1] if len(key) > 1 else ""
        alt = (n % 2 == 1)
        if ra is None or rb is None:
            side = "tm" if rb is None else "tn"
            present = ra if rb is None else rb
            label = sc.side_a if rb is None else sc.side_b
            for li in (0, 1):
                bottom = (li == 1)
                row = [woc(label, side, alt, bottom=bottom),
                       woc("only", side, alt, bottom=bottom),
                       woc(key[0], side, alt, bottom=bottom)]
                for col in _RV_GRID:
                    spec = col[2] if li == 0 else col[5]
                    kind, ref = spec
                    if kind == "pm":
                        text = pmval
                    elif kind == "blank":
                        text = ""
                    elif kind == "tn":
                        text = one.get(ref, "") if (side == "tn" and one) else ""
                    else:
                        text = aval(present, ref)
                    align = lft if (li == 1 and col[4] == "DESCRIPTION") else None
                    row.append(woc(text, side, alt, bottom=bottom, align=align))
                ws.append(row)
            continue
        vals = {(li, ci): value(col[2] if li == 0 else col[5], ra, rb, one)
                for li in (0, 1) for ci, col in enumerate(_RV_GRID)}
        maj = sum(1 for _t, st in vals.values() if st == "hard")
        dif = sum(1 for _t, st in vals.values() if st in ("soft", "hard"))
        for li in (0, 1):
            bottom = (li == 1)
            row = [woc(maj, "hard" if maj else "count", alt, bottom=bottom),
                   woc(dif, "count", alt, bottom=bottom),
                   woc(key[0], "count", alt, bottom=bottom)]
            for ci, col in enumerate(_RV_GRID):
                kind = (col[2] if li == 0 else col[5])[0]
                if kind == "pm":
                    text, st = pmval, "id"
                else:
                    text, st = vals[(li, ci)]
                align = lft if (li == 1 and col[4] == "DESCRIPTION") else None
                row.append(woc(text, st, alt, bottom=bottom, align=align))
            ws.append(row)

    # ---- header merges (aux labels down rows 1-4; group runs across) ----
    for i in range(1, NA_ + 1):
        ws.merged_cells.ranges.add(CellRange(min_col=i, max_col=i, min_row=1, max_row=4))

    def merge_groups(groups, hdr_row):
        i = 0
        while i < NG:
            g, j = groups[i], i
            while j < NG and groups[j] == g:
                j += 1
            if g and j - i > 1:
                ws.merged_cells.ranges.add(CellRange(
                    min_col=NA_ + i + 1, max_col=NA_ + j, min_row=hdr_row, max_row=hdr_row))
            i = j
    merge_groups(g1, 1)
    merge_groups(g2, 3)

    ws.auto_filter.ref = f"A4:{get_column_letter(NC)}{4 + 2 * len(union)}"
    for h, ht in {1: 13, 2: 22, 3: 13, 4: 14}.items():
        ws.row_dimensions[h].height = ht
    WG = {"POST MILE": 11, "PS": 3.5, "LENGTH": 8, "DATE OF REC": 9.5, "H/G": 4,
          "A/C": 4, "ACC-CONT EFF": 10.5, "CITY": 6, "R/U": 4, "EFF-DATE": 10.5,
          "DCR": 12, "LK-AHD": 8, "P": 3.5, "LK-BACK": 8, "CHG/MILE": 13, "DVM": 12}
    for ci, w in {1: 5.5, 2: 5.5, 3: 8}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    for gi, col in enumerate(_RV_GRID):
        lab = col[1]
        ws.column_dimensions[get_column_letter(NA_ + gi + 1)].width = WG.get(lab, 4.6)
    return ws


# --------------------------------------------------------------------------- #
# adapter surface
# --------------------------------------------------------------------------- #
def suggest_name(tsmis_path):
    return suggest_route_name(tsmis_path, "Highway_Detail",
                              "TSMIS_vs_TSN_HighwayDetail")


def _load_pair(tsmis_path, tsn_path):
    """(rows_t, rows_n, warnings) for the shared driver — no input warnings on
    this FLAT pair, so run_compare uses its () default."""
    rows_t, _ = _load_tsmis(tsmis_path)
    rows_n, _ = _load_tsn(tsn_path)
    return rows_t, rows_n, None


def add_report_view(schema, tsmis_path, tsn_path):
    """Augment `schema` with the two-line 'Report View' replica — the printed
    Highway Detail record, comparison-coloured — via the EXISTING
    extra_sheet_writer opt-in plus its Diffs-column self-check. Shared by BOTH
    vs-TSN flavors: the Excel-sourced compare() here and the PDF-sourced
    TSMIS_PDF_VS_TSN in compare_highway_detail_pdf (CMP-AUD-068). The PDF-
    consolidated workbook shares the Excel export's 34-column layout, so the
    replica projects identically no matter which TSMIS render fed the comparison.
    The TSN-only ADT/DCR columns come from the raw TSN file (None for a
    normalized library), read lazily inside the writer, so the workbook is only
    opened when the sheet is actually built (after a successful load).
    `tsmis_path` is accepted for signature parity with the Intersection Detail
    helper; the Highway Detail Report View reads its locations from the compared
    rows, so it isn't needed here."""
    return dataclasses.replace(
        schema,
        extra_sheet_writer=lambda wb, ctx: _write_report_view(
            wb, ctx, _tsn_onesided(Path(tsn_path))),
        report_view_diff_check=("Report View", "B", 2))


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Highway Detail TSMIS-vs-TSN comparison workbook(s). `tsmis_path`
    is the consolidated TSMIS Highway Detail workbook; `tsn_path` the TSN
    statewide (raw or normalized) workbook.

    A per-call schema adds the two-line 'Report View' replica via the EXISTING
    extra_sheet_writer opt-in (the flat Comparison sheet is untouched;
    compare_core stays unmodified). The TSN-only ADT/DCR columns come from the
    raw TSN file (None for a normalized library), read lazily inside the writer."""
    schema = add_report_view(_SCHEMA, tsmis_path, tsn_path)
    return ctc.run_files_compare(
        schema, tsmis_path, tsn_path, out_path,
        banner="Highway Detail Comparison — TSMIS vs TSN", has_route=True,
        loader=_load_pair, deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (openpyxl).",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
