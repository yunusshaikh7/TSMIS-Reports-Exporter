"""Consolidate TSAR: Intersection Summary per-route XLSX into one workbook.

Each per-route export (sheet "Intersection Summary") is a category-count document:
11 blocks (Highway Group, Rural/Urban, Intersection Type, Lighting, Control Types,
Num of Lanes, Mastarm, L/R Channelization, Traffic Flow), each a NUMBER|CODE table.
This consolidator block-walks every route's sheet via the shared, spec-driven
`summary_layout.counts_from_rows` (the SAME mapper the TSN parser uses, so the two
sides can't drift), writing:

  * "Intersection Summary"  one row per route — Route, Total, and one column per
                            canonical category (its compare key as the header).
  * "Combined"              a familiar block-grouped statewide rollup (the source
                            arrangement) with the summed totals.

Console-free (Events sink + ConsolidateResult; no print/input/sys.exit), mirroring
consolidate_ramp_summary. The vs-TSN comparison reads the per-route sheet
(compare_intersection_summary_tsn sums it).
"""
import re
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import summary_layout
from compare_core import is_formula_injection
import outcome
import artifact_store
from events import ConsolidateResult, Events
from paths import (OUTPUT_ROOT, latest_output_day, output_day_dir,
                   stamped_consolidated_filename)

SUBDIR = "intersection_summary"
FILENAME = "tsar_intersection_summary_consolidated.xlsx"
SHEET_NAME = "Intersection Summary"          # per-route source sheet AND output sheet
COMBINED_SHEET = "Combined"
REPORT_NAME = "Intersection Summary"

INPUT_DIR = OUTPUT_ROOT / SUBDIR
OUT_DIR = OUTPUT_ROOT / "consolidated"
OUT_PATH = OUT_DIR / FILENAME

_SPEC = summary_layout.INTERSECTION_SUMMARY_SPEC
_CATS = [c for sec in _SPEC.sections for c in sec.cats]   # every section category, in order
_ROUTE_RE = re.compile(r"Route:\s*(\w+)", re.IGNORECASE)
_TOTAL_RE = re.compile(r"Total Intersections\s*=\s*([\d,]+)", re.IGNORECASE)
_ROUTE_FROM_NAME = re.compile(r"_route_(\w+)\.xlsx$", re.IGNORECASE)


def input_dir_for(day):
    return (output_day_dir(day) / SUBDIR) if day else INPUT_DIR


def out_path_for(day):
    if not day:
        return OUT_PATH
    return output_day_dir(day) / "consolidated" / stamped_consolidated_filename(FILENAME, day)


# --------------------------------------------------------------------------- #
# parse one per-route XLSX -> (route, {slug: count}, total)
# --------------------------------------------------------------------------- #
def parse_route(path):
    """Block-walk one per-route Intersection Summary workbook. Returns
    (route, counts{slug:count}, total_intersections). Route comes from the
    'Route: NNN' header, falling back to the filename."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
        route, total, feed = None, None, []
        for row in ws.iter_rows(values_only=True):
            a = row[0] if len(row) > 0 else None
            b = row[1] if len(row) > 1 else None
            if isinstance(a, (int, float)) and not isinstance(a, bool):
                feed.append((int(a), b))
                continue
            if a is None:
                continue
            s = str(a)
            m = _TOTAL_RE.search(s)
            if m:
                total = int(m.group(1).replace(",", ""))
            m = _ROUTE_RE.search(s)
            if m and route is None:
                route = m.group(1)
            feed.append((None, a))             # header / subheader text
        counts = summary_layout.counts_from_rows(_SPEC, feed)
    finally:
        wb.close()
    if route is None:
        m = _ROUTE_FROM_NAME.search(Path(path).name)
        route = m.group(1) if m else Path(path).stem
    return route, counts, total


def record_has_data(rec):
    """True if a parsed route carries real category figures (not just a route id)."""
    return bool(rec.get("counts")) and sum(rec["counts"].values()) > 0


# --------------------------------------------------------------------------- #
# records -> workbook
# --------------------------------------------------------------------------- #
def _build_combined(wb, statewide, total):
    """A familiar block-grouped statewide rollup sheet (the source arrangement)."""
    ws = wb.create_sheet(COMBINED_SHEET, 0)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", start_color="1F3864")
    sec_fill = PatternFill("solid", start_color="0070C0")
    f_title = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    f_sec = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    f_body = Font(name="Arial", size=10)
    f_total = Font(name="Arial", bold=True, size=12)
    right = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 12
    ws["A1"] = "All Routes Combined — TSAR Intersection Summary"
    ws.merge_cells("A1:B1")
    ws["A1"].font = f_title
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "Total Intersections"
    ws["A2"].font = f_total
    ws["B2"] = total
    ws["B2"].font = f_total
    ws["B2"].alignment = right

    r = 4
    for sec in _SPEC.sections:
        ws.cell(r, 1, sec.name).font = f_sec
        ws.cell(r, 1).fill = sec_fill
        ws.cell(r, 2, "").fill = sec_fill
        r += 1
        for cat in sec.cats:
            ws.cell(r, 1, cat.label).font = f_body
            ws.cell(r, 1).border = border
            c = ws.cell(r, 2, statewide.get(cat.slug, 0))
            c.font = f_body
            c.alignment = right
            c.border = border
            r += 1
        r += 1
    wb.active = wb.index(ws)


def build_workbook(records, out_path, proceed=None):
    """Per-route sheet (Route, Total, one column per category key) + Combined.
    `proceed` (P12) is the pre-replace overwrite gate atomic_save_if evaluates JUST
    BEFORE the os.replace; returns True iff committed (a declined `proceed` keeps the
    prior file)."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header = ["Route", "Total Intersections"] + [c.key for c in _CATS]
    for ci, label in enumerate(header, start=1):
        cell = ws.cell(1, ci, label)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align

    statewide = {c.slug: 0 for c in _CATS}
    for r, rec in enumerate(records, start=2):
        ws.cell(r, 1, rec["route"]).alignment = Alignment(horizontal="left")
        if is_formula_injection(rec["route"]):
            ws.cell(r, 1).data_type = "s"
        ws.cell(r, 2, rec.get("total")).alignment = Alignment(horizontal="right")
        for ci, cat in enumerate(_CATS, start=3):
            v = rec["counts"].get(cat.slug, 0)
            statewide[cat.slug] += v
            ws.cell(r, ci, v).alignment = Alignment(horizontal="center")

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10

    total_all = sum(r.get("total") or 0 for r in records)
    _build_combined(wb, statewide, total_all)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # F9 temp + os.replace + the P12 TOCTOU gate at the replace.
    return artifact_store.atomic_save_if(wb, out_path, proceed or (lambda: True))


# --------------------------------------------------------------------------- #
# entry point
# --------------------------------------------------------------------------- #
def consolidate(events=None, confirm_overwrite=None, day=None,
                input_dir=None, out_path=None):
    """Parse every per-route Intersection Summary XLSX into one workbook.
    Console-free; honors cancel; returns a ConsolidateResult."""
    events = events or Events()
    if not _DEPS_OK:
        return ConsolidateResult(status="error",
                                 message="Required components are missing (openpyxl).")
    confirm = confirm_overwrite or (lambda _p: True)
    day = day or latest_output_day()
    input_dir = input_dir or input_dir_for(day)
    out_path = out_path or out_path_for(day)

    if not input_dir.exists():
        return ConsolidateResult(
            status="error",
            message=(f"The {REPORT_NAME} output folder doesn't exist yet:\n{input_dir}\n\n"
                     f"Export the {REPORT_NAME} report first, then consolidate."))
    files = sorted(input_dir.glob("*.xlsx"))
    files = [f for f in files if not f.name.startswith("~$")]
    if not files:
        return ConsolidateResult(
            status="error",
            message=(f"No {REPORT_NAME} files were found in:\n{input_dir}\n\n"
                     f"Export the {REPORT_NAME} report first, then consolidate."))
    existed_at_confirm = out_path.exists()
    if existed_at_confirm and not confirm(out_path):
        return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"TSAR Intersection Summary Consolidation - {len(files)} file(s)")
    events.on_log("=" * 60)

    records, failed, blank = [], [], []
    for i, p in enumerate(files, 1):
        if events.is_cancelled():
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i:>3}/{len(files)}] {p.name}"
        try:
            route, counts, total = parse_route(str(p))
        except Exception as e:
            events.on_log(f"{prefix} FAILED ({type(e).__name__}): {e}")
            failed.append(p.name)
            continue
        rec = {"route": route, "counts": counts, "total": total}
        if record_has_data(rec):
            records.append(rec)
            events.on_log(f"{prefix} parsed (route {route}, total {total})")
        else:
            events.on_log(f"{prefix} skipped: no intersection data")
            blank.append(p.name)

    if not records:
        return ConsolidateResult(
            status="error",
            message=(f"None of the {len(files)} {REPORT_NAME} file(s) yielded data "
                     f"({len(failed)} failed, {len(blank)} empty). Nothing was written."))

    events.on_log("")
    events.on_log("Writing consolidated workbook...")
    try:
        # P12 TOCTOU: the overwrite gate is INSIDE build_workbook, at the os.replace
        # (atomic_save_if) — a destination that appears during the BUILD is caught.
        committed = build_workbook(records, out_path, proceed=lambda: artifact_store.confirm_late_overwrite(
            out_path, existed_at_confirm, confirm))
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not save {out_path.name}.\n\n"
                     "The file is probably open in Excel. Close it and try again."))
    if not committed:
        return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    incomplete = bool(failed or blank)
    summary_lines = []
    if incomplete:
        summary_lines.append(
            f"⚠ INCOMPLETE — {len(failed) + len(blank)} file(s) left OUT "
            f"({len(failed)} failed, {len(blank)} empty). Re-export before relying on it.")
    summary_lines += [
        f"Parsed:      {len(records)}",
        f"Failed:      {len(failed)} {failed if failed else ''}",
        f"Empty:       {len(blank)} {blank if blank else ''}",
        f"Output file: {out_path}",
    ]
    return ConsolidateResult(status="ok", output_path=str(out_path),
                             summary_lines=summary_lines,
                             completion=outcome.PARTIAL if incomplete else outcome.COMPLETE,
                             skipped_inputs=len(blank), failed_inputs=len(failed))


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
