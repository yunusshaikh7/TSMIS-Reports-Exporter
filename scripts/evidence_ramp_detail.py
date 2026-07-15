"""Ramp Detail adapter for the visual-evidence generator (visual_evidence).

Supplies everything report-specific the engine needs to turn a vs-TSN diff into
a pair of highlighted PDF snippets:

  * the DIFF SOURCE — both comparison sides re-loaded through the comparators'
    OWN loaders/normalizations and compared with compare_core's cell trim, so an
    "example" is exactly a cell the comparison flags, never a re-derivation that
    could drift. The adapter serves BOTH Ramp Detail rows, whose comparisons
    differ: the Excel row rides compare_ramp_detail_tsn (On/Off + Ramp Type are
    context there — never enumerated), the PDF row rides compare_ramp_detail_pdf
    (they're COMPARED — the print carries them, so they enumerate). The
    consolidated workbook names its own source: the PDF-consolidated carries the
    "On/Off" header the Excel-consolidated lacks.
  * the TSMIS locator — the same header-anchored word-window walk the Ramp
    Detail PDF consolidator uses, kept in LOCKSTEP with it (see _locate note),
    but keeping each record's page / y-extent / per-column words so a field
    maps to a cell rectangle.
  * the TSN locator — the STATEWIDE TASAS print (one file, every district;
    per-district files work too) parsed on its fixed column template (header
    anchors pixel-identical across the statewide print; every record is ONE
    line, censused 400/400 against the raw extract), with word positions kept
    so a field maps to a box even when the cell is blank. The whole print is
    indexed ONCE per file (cached on size+mtime).
  * VERIFICATION projections — a candidate is only usable when the value parsed
    back out of each PDF, run through the PDF-flavor's own projections, equals
    the value the comparison compared. The known render skews skip with a
    recorded reason instead of lying: the TSN print TRUNCATES long
    Descriptions; on the EXCEL row the compared Description keeps the
    database's double spaces that the TSMIS print collapses.

Row identity matches the comparison: (route, PM), restricted to keys UNIQUE on
both sides of a route so a highlight is THE row. Console-free;
pdfplumber/openpyxl gated by the engine.
"""
import logging
import re
from collections import defaultdict
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import compare_ramp_detail_pdf as rdp
import compare_ramp_detail_tsn as rd
import consolidate_tsmis_ramp_detail_pdf as rdpdf
from compare_core import _xl_trim
from pdf_table_lib import cluster_by_top
from tsn_load_ramp_detail import SIDECAR_HEADER, tsn_rows_with_dcr  # noqa: F401

log = logging.getLogger("tsmis.evidence")

REPORT_LABEL = "Ramp Detail"
KEY_LABEL = rd.KEY
# Every field either flavor compares (the union): the PDF row also compares the
# print-only On/Off + Ramp Type. Ramp Name and ADT are context in BOTH flavors
# (nothing TSMIS-side to compare them to), so they never enumerate.
_ALWAYS_CONTEXT = ("Ramp Name", "ADT")
FIELDS = [f for f in rd.SHARED_HEADER
          if f != rd.KEY and f not in _ALWAYS_CONTEXT]
# The Excel row's comparison additionally keeps the print-only columns context.
_EXCEL_ROW_SKIPS = ("On/Off", "Ramp Type")

_KEY_I = 1 + rd.KEY_FIELD             # PM's index in a loader row ([route, *header])


def _S(v):
    return ("" if v is None else str(v)).strip()


# --------------------------------------------------------------------------- #
# diff source — both sides through the comparators' own loaders
# --------------------------------------------------------------------------- #
def load_sides(consolidated_path, tsn_path):
    """(tsmis_rows, tsn_rows, meta, note) — both sides in the comparator shape
    ([route, *SHARED_HEADER]). `meta` bundles the (route, key) -> [(district,
    county), ...] sidecar with `pdf` (which flavor the consolidated workbook
    came from — it decides which comparison's projections and compared set
    apply); None when the TSN workbook carries no district info (an old
    normalized library), with `note` saying what to do about it."""
    pdf_sourced = _is_pdf_consolidated(consolidated_path)
    if pdf_sourced:
        tsmis_rows, _ = rdp._load_tsmis_pdf(consolidated_path)
    else:
        tsmis_rows, _ = rd._load_tsmis(consolidated_path)
    tsn_rows, sidecar, note = _load_tsn_with_sidecar(tsn_path)
    if sidecar is None:
        return tsmis_rows, tsn_rows, None, note
    if pdf_sourced:
        # the PDF flavor collapses Description whitespace on BOTH sides
        for r in tsn_rows:
            r[rdp._DESC_I] = rdp._collapse(r[rdp._DESC_I])
    return tsmis_rows, tsn_rows, {"dc": sidecar, "pdf": pdf_sourced}, None


def _is_pdf_consolidated(path):
    """True when the consolidated workbook is the PDF-sourced one (it carries
    the print-only "On/Off" header the Excel export lacks)."""
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb[rd.TSMIS_SHEET] if rd.TSMIS_SHEET in wb.sheetnames else wb.active
        header = next(ws.iter_rows(values_only=True), []) or []
        return "On/Off" in [(_S(c)) for c in header]
    finally:
        wb.close()


def _load_tsn_with_sidecar(path):
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if rd.NORMALIZED_SHEET in wb.sheetnames:
            it = wb[rd.NORMALIZED_SHEET].iter_rows(values_only=True)
            header = [_S(c) for c in (next(it, []) or [])]
            # v4 gate (CMP-AUD-045): a pre-v4 library lacks the District column
            # / PM-Suffix sidecar the county-aware identity needs — same rule as
            # the comparison loader, surfaced as the evidence rebuild note.
            if "District" not in header or "TSN PM Suffix" not in header:
                return [], None, ("the normalized TSN library predates the "
                                  "county-aware evidence columns — rebuild the "
                                  "TSN library (Settings) and run the "
                                  "comparison again")
            rows, sidecar = [], defaultdict(list)
            for r in it:
                if not r or all(c in (None, "") for c in r):
                    continue
                row = rd._normalized_row(r)    # re-projected like the comparison
                rows.append(row)
                dist = _S(r[1 + len(rd.SHARED_HEADER)])
                cnty = _S(r[2 + len(rd.SHARED_HEADER)]).rstrip(".")
                sidecar[(row[0], row[_KEY_I])].append((dist, cnty))
            return rows, sidecar, None
    finally:
        wb.close()
    rows, dcr = tsn_rows_with_dcr(path)          # a RAW statewide file
    sidecar = defaultdict(list)
    for row, (dist, cnty, _sfx) in zip(rows, dcr):
        sidecar[(row[0], row[_KEY_I])].append((dist, cnty))
    return rows, sidecar, None


def enumerate_diffs(tsmis_rows, tsn_rows, meta):
    """{field: [example]} over (route, PM) keys UNIQUE per route on BOTH sides —
    each example a cell the ROW'S OWN comparison flags (the Excel row's flavor
    never compares the print-only columns, so they never enumerate there).
    Equality mirrors compare_core's cell compare (loader projections + the
    Excel TRIM), so inequality here == a red cell there."""
    sidecar, pdf_sourced = meta["dc"], meta["pdf"]
    skips = () if pdf_sourced else _EXCEL_ROW_SKIPS
    a_route, b_route = defaultdict(list), defaultdict(list)
    for r in tsmis_rows:
        a_route[r[0]].append(r)
    for r in tsn_rows:
        b_route[r[0]].append(r)
    diffs = defaultdict(list)
    for route in sorted(set(a_route) & set(b_route)):
        a_ct, b_ct = defaultdict(int), defaultdict(int)
        for r in a_route[route]:
            a_ct[r[_KEY_I]] += 1
        for r in b_route[route]:
            b_ct[r[_KEY_I]] += 1
        a_by = {r[_KEY_I]: r for r in a_route[route] if a_ct[r[_KEY_I]] == 1}
        b_by = {r[_KEY_I]: r for r in b_route[route] if b_ct[r[_KEY_I]] == 1}
        for key in set(a_by) & set(b_by):
            ra, rb = a_by[key], b_by[key]
            for i, f in enumerate(rd.SHARED_HEADER):
                if f == rd.KEY or f in _ALWAYS_CONTEXT or f in skips:
                    continue
                va, vb = _xl_trim(ra[1 + i]), _xl_trim(rb[1 + i])
                if va != vb:
                    dist, cnty = (sidecar.get((route, key)) or [("", "")])[0]
                    # The engine's locators key on the plain normalized-PM text
                    # (+ county for TSN); the D4 PhysicalKey's str payload IS
                    # that text (CMP-AUD-045).
                    diffs[f].append(dict(route=route, key=str(key), field=f,
                                         va=va, vb=vb, dist=dist, cnty=cnty))
    return diffs


# --------------------------------------------------------------------------- #
# verification projections — the PDF flavor's own, per field
# --------------------------------------------------------------------------- #
def project(field, raw):
    """A raw PDF token -> the compared form, via the PDF flavor's projections
    (null-render tokens to blank, the print's N -> TSN's O, Description prefix
    strip + whitespace collapse) + compare_core's cell trim."""
    v = rd._v(raw)
    if field == "District":
        v = rd._dist_cnty(raw)[0]
    elif field == "Date of Record":
        v = rd._iso_date(raw)
    elif field == "Description":
        v = rdp._collapse(rd._strip_desc_prefix(raw))
        if v == rdp._NULL_DESC:
            v = ""
    elif field in ("Area 4", "On/Off"):
        v = rdp._null_blank(v)
        if field == "On/Off" and v == "N":
            v = "O"
    return _xl_trim(v)


# --------------------------------------------------------------------------- #
# TSMIS side — the per-route "TSAR: Ramp Detail (PDF)" export
# --------------------------------------------------------------------------- #
def tsmis_pdf_path(pdf_dir, route):
    return Path(pdf_dir) / f"tsar_ramp_detail_route_{route}.pdf"


# shared field -> the consolidator's column key (see rdpdf._COL_ORDER).
# District (CMP-AUD-185) lives inside the print's Location column on both PDFs.
_TSMIS_COL = {"PR": "pr", "District": "loc", "Date of Record": "date", "HG": "hg",
              "Area 4": "area4", "City Code": "city", "R/U": "ru",
              "Description": "desc", "On/Off": "onoff", "Ramp Type": "rtype"}
# ... and the window each column spans (boundary names in rdpdf's dict), for
# boxing a BLANK cell.
_COL_WIN = {"pr": ("loc_pr", "pr_pm"), "date": ("pm_date", "date_hg"),
            "hg": ("date_hg", "hg_area"), "area4": ("hg_area", "area_city"),
            "city": ("area_city", "city_ru"), "ru": ("city_ru", "ru_onoff"),
            "onoff": ("ru_onoff", "onoff_type"), "rtype": ("onoff_type", "type_desc")}


def locate_tsmis(pdf_path, needed_keys):
    """{normalized_pm: [record]} for `needed_keys`; a record carries the parsed
    13-column row plus geometry (page, y-extent, per-column words, boundaries).

    LOCKSTEP: this walk mirrors consolidate_tsmis_ramp_detail_pdf.parse_pdf
    step for step (the per-page header anchors, the banner/header skip, the
    PM-window row test, the desc-fragment attach) — it only ADDS position
    capture. A behavior change there must land here too; check_visual_evidence
    pins the shared pieces so a drift fails the gate."""
    found = defaultdict(list)
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            words = page.extract_words()
            b, hdr_bottom = rdpdf._page_header(words)
            if b is None:
                continue
            page_rows, frags = [], []
            for top, line_words in rdpdf._cluster_lines(words):
                if top <= hdr_bottom + 2:
                    continue
                cols = {k: [] for k in rdpdf._COL_ORDER}
                for w in line_words:
                    xc = (w["x0"] + w["x1"]) / 2
                    for name, (lo_k, hi_k) in _COL_WIN.items():
                        if b[lo_k] <= xc < b[hi_k]:
                            cols[name].append(w)
                            break
                    else:
                        cols["desc" if xc >= b["type_desc"] else "loc"].append(w)
                vals = rdpdf._classify_words(line_words, b)
                if rdpdf.PM_RE.fullmatch(vals["pm"]):
                    page_rows.append({"top": top, "vals": vals, "cols": cols,
                                      "bottom": max(w["bottom"] for w in line_words),
                                      "words": line_words})
                elif vals["desc"] and not any(
                        vals[k] for k in rdpdf._COL_ORDER if k != "desc"):
                    frags.append((top, vals["desc"], line_words))
            parts = {id(pr): [(pr["top"], pr["vals"]["desc"])] for pr in page_rows}
            for ftop, ftext, fwords in frags:
                best = min(page_rows, key=lambda pr: abs(pr["top"] - ftop),
                           default=None)
                if best is None or abs(best["top"] - ftop) > rdpdf.FRAG_MAX_DIST:
                    continue
                parts[id(best)].append((ftop, ftext))
                best["cols"]["desc"].extend(fwords)
                best["top"] = min(best["top"], ftop)
                best["bottom"] = max(best["bottom"],
                                     max(w["bottom"] for w in fwords))
                best["words"] = best["words"] + list(fwords)
            for pr in page_rows:
                pr["vals"]["desc"] = rdpdf.join_desc_parts(
                    [t for _, t in sorted(parts[id(pr)])])
                key = rd._norm_pm(pr["vals"]["pm"])
                if key in needed_keys:
                    found[key].append({"vals": pr["vals"], "cols": pr["cols"],
                                       "b": b, "page": page_no,
                                       "top": pr["top"], "bottom": pr["bottom"],
                                       "words": pr["words"]})
    return found


def tsmis_value(rec, field):
    """The compared value this PDF record carries for `field` (verification)."""
    return project(field, rec["vals"][_TSMIS_COL[field]])


def tsmis_box(rec, field):
    """(page_no, cell_box, record_yspan, record_xspan) for `field`'s cell. A
    blank cell boxes its column window — on a header-anchored layout the window
    IS the cell."""
    col = _TSMIS_COL[field]
    hits = rec["cols"].get(col) or []
    if hits:
        x0 = min(w["x0"] for w in hits)
        x1 = max(w["x1"] for w in hits)
    elif col in _COL_WIN:
        lo_k, hi_k = _COL_WIN[col]
        x0, x1 = rec["b"][lo_k] + 2, rec["b"][hi_k] - 2
    else:
        return None                              # desc/loc always have words
    xspan = (min(w["x0"] for w in rec["words"]) - 4,
             max(w["x1"] for w in rec["words"]) + 4)
    return (rec["page"], (x0 - 2, rec["top"] - 2, x1 + 2, rec["bottom"] + 2),
            (rec["top"], rec["bottom"]), xspan)


# --------------------------------------------------------------------------- #
# TSN side — the statewide TASAS print (fixed column template)
# --------------------------------------------------------------------------- #
# One fixed template document-wide (header anchors pixel-identical across the
# 500-page statewide print; field extents censused on 415 order-assignable full
# rows, then 400/400 sampled records verified value-identical to the raw
# extract). Every record is ONE line; long Descriptions TRUNCATE (a known
# render skew — those examples skip with a reason). Words are assigned to
# windows by MAX OVERLAP, like the Intersection Detail print.
_L_WIN = [
    ("LOC", 4, 70), ("PR", 80, 91.8), ("PM", 91.8, 136), ("DATE_REC", 136, 209),
    ("HG", 209, 241), ("AREA4", 241, 262), ("CITY", 262, 299), ("RU", 299, 317),
    ("ONOFF", 317, 335), ("ADT_YR", 335, 368), ("ADT", 368, 418),
    ("TYPE", 418, 442), ("EFF_DATE", 442, 513), ("DESC", 513, 800),
]
# shared comparison field -> window name.
TSN_CELL = {"PR": "PR", "District": "LOC", "Date of Record": "DATE_REC",
            "HG": "HG", "Area 4": "AREA4", "City Code": "CITY", "R/U": "RU",
            "On/Off": "ONOFF", "Ramp Type": "TYPE", "Description": "DESC"}
_PM_LINE_RE = re.compile(r"^\d{3}\.\d{3}$")
_LOC_RE = re.compile(r"^(\d{2})-([A-Z]{1,4}\.?)-(\w+)$")

_WORD_GAP = 1.6
# One statewide print is ~500 pages of word extraction — index each file ONCE
# and serve every district's locate from the cache (keyed on size+mtime; a few
# entries so per-district drops work too).
_INDEX_CACHE = {}
_INDEX_CACHE_MAX = 4


def _words_of(chars):
    ws, cur, last = [], None, None
    for ch in sorted(chars, key=lambda c: c["x0"]):
        if last is None or ch["x0"] - last > _WORD_GAP:
            if cur:
                ws.append(cur)
            cur = {"t": ch["text"], "x0": ch["x0"], "x1": ch["x1"]}
        else:
            cur["t"] += ch["text"]
            cur["x1"] = ch["x1"]
        last = ch["x1"]
    if cur:
        ws.append(cur)
    return ws


def _assign_win(words):
    """{window name: (joined text, [word boxes])} by MAX x-overlap per word."""
    hit = {name: [] for name, _lo, _hi in _L_WIN}
    for w in words:
        best, best_ov = None, 0.0
        for name, lo, hi in _L_WIN:
            ov = min(w["x1"], hi) - max(w["x0"], lo)
            if ov > best_ov:
                best, best_ov = name, ov
        if best is not None:
            hit[best].append(w)
    return {name: (" ".join(w["t"] for w in ws), ws) for name, ws in hit.items()}


def _print_index(path, events=None):
    """Parse one TSN Ramp Detail print into {(county, route, pm): [record]} +
    the set of districts it covers. Cached on (size, mtime) — the statewide
    file is indexed once, then every district's locate is a lookup."""
    path = Path(path)
    st = path.stat()
    sig = (st.st_size, st.st_mtime_ns)
    ent = _INDEX_CACHE.get(str(path))
    if ent and ent["sig"] == sig:
        return ent
    records = defaultdict(list)
    districts = set()
    with pdfplumber.open(path) as pdf:
        n_pages = len(pdf.pages)
        for pi, page in enumerate(pdf.pages, 1):
            if events is not None and pi % 200 == 0:
                events.on_log(f"    …TSN print {path.name}: page {pi}/{n_pages}")
            page_chars = [c for c in page.chars if c["text"].strip()]
            # a statewide print is ~500 pages in ONE handle — without this,
            # pdfplumber's per-page caches accumulate to gigabytes by the end.
            page.flush_cache()
            for _top, chars in cluster_by_top(page_chars, 3):
                words = _words_of(chars)
                if not (words and _LOC_RE.match(words[0]["t"])):
                    continue
                a = _assign_win(words)
                if not _PM_LINE_RE.match(a["PM"][0]):
                    continue
                m = _LOC_RE.match(words[0]["t"])
                dist, cnty = m.group(1), m.group(2).rstrip(".")
                route = rd._norm_route(m.group(3))
                districts.add(dist)
                key = (cnty, route, rd._norm_pm(a["PM"][0]))
                records[key].append({
                    "a": a, "page": pi, "dist": dist,
                    "top": min(c["top"] for c in chars),
                    "bottom": max(c["bottom"] for c in chars)})
    ent = {"sig": sig, "districts": districts, "records": dict(records)}
    while len(_INDEX_CACHE) >= _INDEX_CACHE_MAX:
        _INDEX_CACHE.pop(next(iter(_INDEX_CACHE)))
    _INDEX_CACHE[str(path)] = ent
    return ent


def district_index(pdf_dir, events=None):
    """{district('01'..'12'): path} by INDEXING each PDF and reading which
    districts its own records cover (filenames are the user's business; the
    single statewide print maps every district to itself). The heavy word
    extraction happens here, once per file — locate_tsn is lookups after."""
    index = {}
    for p in sorted(Path(pdf_dir).glob("*.pdf")):
        try:
            ent = _print_index(p, events)
        except Exception as e:                            # unreadable/odd PDF
            log.warning("evidence: %s unreadable: %s: %s",
                        p.name, type(e).__name__, e)
            if events:
                events.on_log(f"    note: {p.name} unreadable, skipped")
            continue
        if not ent["districts"]:
            log.info("evidence: %s has no Ramp Detail records; skipped", p.name)
            continue
        for dist in ent["districts"]:
            index.setdefault(dist, p)
    return index


def locate_tsn(pdf_path, needed_routes, needed_keys):
    """{(county, route, pm): [record]} served from the file's cached index."""
    del needed_routes                       # the key filter is already exact
    ent = _print_index(pdf_path)
    return {k: ent["records"][k] for k in needed_keys if k in ent["records"]}


def tsn_value(rec, field):
    """The compared value this print record carries for `field`."""
    return project(field, rec["a"][TSN_CELL[field]][0])


def tsn_box(rec, field):
    """(page_no, cell_box, record_yspan, record_xspan) for `field`'s cell. A
    blank cell boxes its fixed template window — on a fixed template the
    window IS the cell."""
    name = TSN_CELL[field]
    assigned = rec["a"][name][1]
    if assigned:
        x0, x1 = min(w["x0"] for w in assigned), max(w["x1"] for w in assigned)
    else:
        lo, hi = next((lo, hi) for n, lo, hi in _L_WIN if n == name)
        x0, x1 = lo + 1, hi - 3
    words = [w for _n, (_t, ws) in rec["a"].items() for w in ws]
    xspan = (min(w["x0"] for w in words) - 4, max(w["x1"] for w in words) + 4)
    return (rec["page"], (x0 - 2, rec["top"] - 2, x1 + 2, rec["bottom"] + 2),
            (rec["top"], rec["bottom"]), xspan)
