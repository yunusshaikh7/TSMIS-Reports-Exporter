"""Cross-environment comparison MATRIX engine (Everything tab).

A thin ORCHESTRATION layer over the already-audited cross-environment comparison
(`compare_env`) and the always-current Export-Everything store. It NEVER
recomputes report contents and NEVER touches `compare_core` — it only:

  * enumerates the report x environment cells from the store,
  * computes per-cell EXPORT freshness (newest file mtime) via report_library,
  * computes per-cell COMPARISON freshness (the comparison workbook's mtime vs
    the two source folders' newest export mtime),
  * orchestrates `compare_env.<adapter>.compare_folders(...)` to (re)build a
    cell's comparison of (report, env) against the BASELINE env, writing into
    <dest>/comparisons/<baseline>/, and
  * caches each comparison's verdict + discrepancy counts in a sidecar so the
    snapshot (which the GUI renders) stays a pure, offline filesystem read.

Console-free like the rest of the core: progress is reported through the Events
sink and exceptions are raised — never print/input/sys.exit. Only gui_api /
gui_worker drive it.
"""
import json
import logging
import os
import shutil
import time
from pathlib import Path

import report_library
import reports
from common import DATA_SOURCES, ENVIRONMENTS

log = logging.getLogger("tsmis.matrix")

BASELINE_DEFAULT = "ssor-prod"
COMPARISONS_DIRNAME = "comparisons"
_RESULTS_FILE = "_results.json"
_NEQ = " ≠ "                       # the compare_core diff marker (read-only here)
_MTIME_TOL_S = 1.0                 # float-mtime equality tolerance for the cache


# --------------------------------------------------------------------------- #
# cell identity
# --------------------------------------------------------------------------- #
def env_keys():
    """The matrix columns: every data-source/environment combo, in display order
    (ssor before ars; prod/test/dev)."""
    return [f"{s}-{e}" for s in DATA_SOURCES for e in ENVIRONMENTS]


def default_env_label(key):
    """A readable column label from an env key, with no dependency on the report
    registry: 'ssor-prod' -> 'SSOR / Prod'."""
    src, _, env = str(key).partition("-")
    return f"{src.upper()} / {env.title()}" if env else str(key).upper()


def _row_defs():
    """{row_key: (label, subdir, export_idx, adapter, has_route)} from the
    registry. has_route: the XLSX reports compare in the consolidated (Route +
    columns) shape; Ramp Summary (PDF, no sheet_name) is per-route."""
    out = {}
    for row_key, label, subdir, export_idx, adapter in reports.matrix_rows():
        has_route = getattr(adapter, "sheet_name", None) is not None
        out[row_key] = (label, subdir, export_idx, adapter, has_route)
    return out


# --------------------------------------------------------------------------- #
# paths (decision b: <dest>/comparisons/<baseline>/...) — STABLE, dateless names
# --------------------------------------------------------------------------- #
def comparisons_root(dest, baseline_key):
    return Path(dest) / COMPARISONS_DIRNAME / baseline_key


def comparison_path(dest, baseline_key, row_key, cell_key):
    """The comparison workbook for (row, cell-env) under one baseline.

    DELIBERATELY a deterministic, dateless name (NOT adapter.suggest_name, which
    embeds today's date): the matrix overwrites in place and uses the file's
    MTIME as the freshness signal, so the name must be stable across days. Do not
    'fix' this to a dated name."""
    return comparisons_root(dest, baseline_key) / f"{cell_key}_{row_key}.xlsx"


def _results_path(dest, baseline_key):
    return comparisons_root(dest, baseline_key) / _RESULTS_FILE


# --------------------------------------------------------------------------- #
# the comparison-result cache (verdict + discrepancy counts), keyed by baseline
# --------------------------------------------------------------------------- #
def load_results(dest, baseline_key):
    """{row_key: {cell_key: {verdict, diff_cells, one_sided, built_at_mtime}}}.
    Tolerant: missing/corrupt -> {} (never raises)."""
    p = _results_path(dest, baseline_key)
    try:
        with open(p, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except OSError:
        return {}                            # not written yet (first run) — expected
    except ValueError as e:                  # corrupt JSON: surface it, then degrade
        log.warning("matrix: corrupt results cache %s (%s: %s); treating as empty",
                    p, type(e).__name__, e)
        return {}


def _save_results(dest, baseline_key, data):
    p = _results_path(dest, baseline_key)
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_name(p.name + ".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f)
        os.replace(tmp, p)
    except OSError as e:
        log.warning("matrix: could not write results cache %s: %s: %s",
                    p, type(e).__name__, e)


def record_result(dest, baseline_key, row_key, cell_key, verdict,
                  diff_cells, one_sided, built_at_mtime):
    data = load_results(dest, baseline_key)
    data.setdefault(row_key, {})[cell_key] = {
        "verdict": verdict, "diff_cells": diff_cells,
        "one_sided": one_sided, "built_at_mtime": built_at_mtime,
    }
    _save_results(dest, baseline_key, data)


# --------------------------------------------------------------------------- #
# read discrepancy counts back from a produced VALUES workbook (no COM/F9)
# --------------------------------------------------------------------------- #
def read_counts(values_path, has_route):
    """(diff_cells, one_sided) read straight off a VALUES-flavor comparison
    workbook's Comparison sheet content: cells carrying the ' ≠ ' marker, and
    non-'Both' statuses. The values flavor stores literal results, so no Excel
    recalc is needed. Returns (None, None) if unreadable. Layout:
      has_route:  A Route | B key | C # | D A-Row | E B-Row | F Status | G Diffs | H.. fields
      flat:       A Route | B # | C A-Row | D B-Row | E Status | F Diffs | G.. fields
    """
    status_col = 6 if has_route else 5          # 1-based
    first_field = 8 if has_route else 7
    try:
        from openpyxl import load_workbook
        wb = load_workbook(values_path, read_only=True, data_only=True)
    except Exception as e:                       # noqa: BLE001 (best-effort read)
        log.debug("read_counts: can't open %s (%s: %s)", values_path,
                  type(e).__name__, e)
        return (None, None)
    try:
        ws = wb["Comparison"]
        one_sided = diff_cells = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v is None for v in row):
                continue
            status = row[status_col - 1] if len(row) >= status_col else None
            if status and status != "Both":
                one_sided += 1
            for ci in range(first_field - 1, len(row)):
                v = row[ci]
                if isinstance(v, str) and _NEQ in v:
                    diff_cells += 1
        return (diff_cells, one_sided)
    except Exception as e:                       # noqa: BLE001
        log.debug("read_counts: can't read %s (%s: %s)", values_path,
                  type(e).__name__, e)
        return (None, None)
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# per-cell comparison freshness (decision c: mtime staleness)
# --------------------------------------------------------------------------- #
def comparison_state(dest, baseline_key, row_key, cell_key, subdir,
                     cell_ages_map, results):
    """{built, mtime, stale, reason, missing_side, verdict, diff_cells,
    one_sided} for one non-baseline cell. STALE when the comparison file is
    missing, or either side's export mtime is newer than it. The cached
    verdict/counts are surfaced only when the cache's recorded mtime still
    matches the file (else they read as unknown and the cell shows 're-run')."""
    base_m = cell_ages_map.get(baseline_key, {}).get(subdir, {}).get("mtime")
    cell_m = cell_ages_map.get(cell_key, {}).get(subdir, {}).get("mtime")
    if base_m is None and cell_m is None:
        missing_side = "both"
    elif base_m is None:
        missing_side = "baseline"
    elif cell_m is None:
        missing_side = "cell"
    else:
        missing_side = None

    try:
        cmp_m = comparison_path(dest, baseline_key, row_key, cell_key).stat().st_mtime
    except OSError:
        cmp_m = None
    built = cmp_m is not None

    verdict = diff_cells = one_sided = None
    rec = (results.get(row_key, {}) or {}).get(cell_key)
    if built and rec and abs(float(rec.get("built_at_mtime", -1)) - cmp_m) < _MTIME_TOL_S:
        verdict = rec.get("verdict")
        diff_cells = rec.get("diff_cells")
        one_sided = rec.get("one_sided")

    if not built:
        stale, reason = True, "missing"
    else:
        newer_base = base_m is not None and base_m > cmp_m + _MTIME_TOL_S
        newer_cell = cell_m is not None and cell_m > cmp_m + _MTIME_TOL_S
        if newer_base and newer_cell:
            stale, reason = True, "both_newer"
        elif newer_base:
            stale, reason = True, "baseline_newer"
        elif newer_cell:
            stale, reason = True, "cell_newer"
        else:
            stale, reason = False, "fresh"
    return {"supported": True, "built": built, "mtime": cmp_m, "stale": stale,
            "reason": reason, "missing_side": missing_side, "verdict": verdict,
            "diff_cells": diff_cells, "one_sided": one_sided}


# --------------------------------------------------------------------------- #
# TSN / cross-format modes. The matrix only ORCHESTRATES the existing (manual)
# comparison adapters — compare_env / compare_highway_log / compare_highway_log_pdf
# are NEVER edited. TSN drops: <dest>/_tsn_input/<subdir>/; the TSN + cross-format
# comparison sheets: <dest>/comparisons/tsn/ (apart from the cross-env tree).
# --------------------------------------------------------------------------- #
TSN_INPUT_DIRNAME = "_tsn_input"
TSN_COMPARE_DIRNAME = "tsn"            # under comparisons/
_TSN_RESULTS_FILE = "_tsn_results.json"


def _safe_mtime(p):
    try:
        return Path(p).stat().st_mtime
    except OSError:
        return None


def tsn_input_root(dest, subdir):
    """Where the user drops the TSN dataset (a consolidated workbook or the
    district PDFs). TSN is ONE dataset per report family, so both Highway Log TSN
    modes (Excel-vs-TSN, PDF-vs-TSN) share <dest>/_tsn_input/highway_log/."""
    return Path(dest) / TSN_INPUT_DIRNAME / subdir


def tsn_comparisons_root(dest):
    return Path(dest) / COMPARISONS_DIRNAME / TSN_COMPARE_DIRNAME


def tsn_source(dest, subdir, selected_file=None):
    """Resolve the TSN dataset for a report `subdir`, returning {kind:
    file|consolidated|pdfs|raw|none, path?, mtime?, pdf_count?, raw_count?}.

    Delegates to the canonical TSN library (tsn_library.resolve): an explicit
    user-picked `selected_file` wins; else the library's consolidated workbook;
    else its raw file(s) -> the 'consolidate first' state; else the legacy
    dest-scoped drop <dest>/_tsn_input/<subdir>/ and the global legacy locations
    (back-compat) — so an existing install keeps resolving until imported."""
    import tsn_library                              # lazy: no import cycle
    return tsn_library.resolve(subdir, selected_file, legacy_dest=dest)


# --- per-row comparison MODE registry -------------------------------------- #
# A row's "env" (cross-environment) mode is supported only when it has a
# compare_env folder adapter. Highway Log is TWO rows (Excel + PDF), each with its
# own TSN + cross-format modes; the other reports get one greyed "tsn" placeholder
# (no comparison code yet). A mode dict: id, label, kind ("env"|"tsn"|"self"),
# supported, env_subdir, plus tsn_subdir+fmt (tsn) or other_subdir (self).
def _row_modes(row_key, subdir, adapter):
    env = {"id": "env", "label": "Cross-environment", "kind": "env",
           "supported": adapter is not None, "env_subdir": subdir}
    if row_key == "highway_log":            # TSMIS Excel
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn", "supported": True,
                 "env_subdir": "highway_log", "tsn_subdir": "highway_log", "fmt": "excel"},
                {"id": "vs_pdf", "label": "vs TSMIS PDF", "kind": "self", "supported": True,
                 "env_subdir": "highway_log", "other_subdir": "highway_log_pdf"}]
    if row_key == "highway_log_pdf":        # TSMIS PDF (cross-env not coded -> greyed)
        return [env,
                {"id": "tsn", "label": "vs TSN", "kind": "tsn", "supported": True,
                 "env_subdir": "highway_log_pdf", "tsn_subdir": "highway_log", "fmt": "pdf"},
                {"id": "vs_excel", "label": "vs TSMIS Excel", "kind": "self", "supported": True,
                 "env_subdir": "highway_log_pdf", "other_subdir": "highway_log"}]
    return [env,
            {"id": "tsn", "label": "vs TSN", "kind": "tsn",
             "supported": tsn_supported(row_key),
             "env_subdir": subdir, "tsn_subdir": subdir, "fmt": None}]


def _mode_by_id(modes, mode_id):
    for m in modes:
        if m["id"] == mode_id:
            return m
    return modes[0]                          # default to env


def tsn_comparator_for(row_key):
    """The vs-TSN comparison adapter for a report row (a `"files"` module/instance
    exposing compare(path_a, path_b, out_path, …)), or None when the report has no
    TSN comparator yet. The single registry of which reports support vs-TSN — adding
    a report here (+ its CompareSchema module) flips it on in BOTH matrices."""
    if row_key == "highway_log":
        import compare_highway_log as _m            # lazy
        return _m
    if row_key == "highway_log_pdf":
        import compare_highway_log_pdf as _m
        return _m.TSMIS_PDF_VS_TSN
    if row_key == "ramp_detail":
        import compare_ramp_detail_tsn as _m
        return _m
    if row_key == "ramp_summary":
        import compare_ramp_summary_tsn as _m       # AGGREGATE recipe
        return _m
    if row_key == "intersection_summary":
        import compare_intersection_summary_tsn as _m   # AGGREGATE
        return _m
    if row_key == "intersection_detail":
        import compare_intersection_detail_tsn as _m    # FLAT
        return _m
    if row_key == "highway_sequence":
        import compare_highway_sequence_tsn as _m       # FLAT (county+PM key)
        return _m
    return None


def tsn_supported(row_key):
    """True when the report row has a coded vs-TSN comparator (the matrices show it
    live). Every report qualifies as of v0.17.0; an uncoded row would grey out —
    defensive, kept for any future report added before its comparator lands."""
    return tsn_comparator_for(row_key) is not None


def tsn_subdir_for(row_key, subdir, adapter=None):
    """The TSN dataset key (per-row `tsn_subdir`) for a row's vs-TSN comparison.
    Both Highway Log rows share ONE TSN dataset ('highway_log'); every other report
    uses its own subdir. The single source of truth is the row's 'tsn' mode in
    _row_modes — this is what replaces day_matrix's hardcoded TSN_SUBDIR."""
    for m in _row_modes(row_key, subdir, adapter):
        if m.get("kind") == "tsn":
            return m.get("tsn_subdir") or subdir
    return subdir


def tsn_capable(row_key):
    """True if the row offers a coded comparison beyond cross-environment (the two
    Highway Log rows). Drives the 'tsn_capable' chip hint."""
    defs = _row_defs()
    if row_key not in defs:
        return False
    _l, subdir, _i, adapter, _hr = defs[row_key]
    return any(m["kind"] != "env" and m["supported"]
               for m in _row_modes(row_key, subdir, adapter))


# --- comparison-output paths + the non-env (TSN/self) counts cache --------- #
def mode_out_path(dest, baseline_key, row_key, cell_key, mode):
    """Where a cell's comparison workbook lives. Cross-env stays under
    comparisons/<baseline>/ (existing); TSN/self modes go to comparisons/tsn/ with
    the mode in the name so flavors coexist."""
    if mode["kind"] == "env":
        return comparison_path(dest, baseline_key, row_key, cell_key)
    return tsn_comparisons_root(dest) / f"{cell_key}_{row_key}_{mode['id']}.xlsx"


def out_path_for_cell(dest, baseline_key, row_key, cell_key, mode_id, row_defs=None):
    """The comparison workbook path for one cell under the row's mode — for the
    'open' action. None for an unknown row."""
    rows = row_defs if row_defs is not None else _row_defs()
    if row_key not in rows:
        return None
    _l, subdir, _i, adapter, _hr = rows[row_key]
    mode = _mode_by_id(_row_modes(row_key, subdir, adapter), mode_id)
    return mode_out_path(dest, baseline_key, row_key, cell_key, mode)


def _tsn_results_path(dest):
    return tsn_comparisons_root(dest) / _TSN_RESULTS_FILE


def load_tsn_results(dest):
    """{ "<row>|<mode>": {cell_key: {verdict, diff_cells, one_sided,
    built_at_mtime}} } — the non-env (TSN/self) comparison counts cache."""
    p = _tsn_results_path(dest)
    try:
        with open(p, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except OSError:
        return {}                            # not written yet (first run) — expected
    except ValueError as e:                  # corrupt JSON: surface it, then degrade
        log.warning("matrix: corrupt TSN results cache %s (%s: %s); treating as empty",
                    p, type(e).__name__, e)
        return {}


def record_tsn_result(dest, result_key, cell_key, verdict, diff_cells, one_sided,
                      built_at_mtime):
    data = load_tsn_results(dest)
    data.setdefault(result_key, {})[cell_key] = {
        "verdict": verdict, "diff_cells": diff_cells,
        "one_sided": one_sided, "built_at_mtime": built_at_mtime,
    }
    p = _tsn_results_path(dest)
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_name(p.name + ".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f)
        os.replace(tmp, p)
    except OSError as e:
        log.warning("matrix: could not write TSN results cache %s: %s: %s",
                    p, type(e).__name__, e)


# --- unified per-cell comparison state ------------------------------------- #
def _cmp_state(out_path, sources, rec):
    """{supported, built, mtime, stale, reason, missing_side, verdict, diff_cells,
    one_sided} for one comparison cell. `sources` is the input sides
    [{name, present, mtime}]; STALE when the workbook is missing or any present
    source is newer than it."""
    missing = [s["name"] for s in sources if not s.get("present")]
    missing_side = missing[0] if missing else None
    cmp_m = _safe_mtime(out_path)
    built = cmp_m is not None
    verdict = diff_cells = one_sided = None
    if built and rec and abs(float(rec.get("built_at_mtime", -1)) - cmp_m) < _MTIME_TOL_S:
        verdict = rec.get("verdict")
        diff_cells = rec.get("diff_cells")
        one_sided = rec.get("one_sided")
    if not built:
        stale, reason = True, "missing"
    else:
        newer = [s["name"] for s in sources
                 if s.get("mtime") is not None and s["mtime"] > cmp_m + _MTIME_TOL_S]
        stale = bool(newer)
        reason = ("both_newer" if len(newer) > 1
                  else (f"{newer[0]}_newer" if newer else "fresh"))
    return {"supported": True, "built": built, "mtime": cmp_m, "stale": stale,
            "reason": reason, "missing_side": missing_side, "verdict": verdict,
            "diff_cells": diff_cells, "one_sided": one_sided}


# --------------------------------------------------------------------------- #
# the snapshot the GUI renders (pure filesystem read)
# --------------------------------------------------------------------------- #
def apply_order(keys, order):
    """Reorder `keys` by the user's preference list `order`: keys named in `order`
    come first (in that order, only those actually present), then any remaining keys
    in their original order. `order` is a PREFERENCE, not a filter — unknown/stale
    keys in `order` are ignored, and keys missing from `order` are kept (appended in
    their natural order), so a report/env added or removed later degrades gracefully."""
    if not order:
        return list(keys)
    keyset = set(keys)
    front = [k for k in order if k in keyset]
    frontset = set(front)
    return front + [k for k in keys if k not in frontset]


def matrix_snapshot(dest, baseline_key=BASELINE_DEFAULT, envs=None,
                    env_labels=None, row_defs=None, hidden=None, hidden_envs=None,
                    row_modes=None, tsn_files=None, now=None,
                    row_order=None, env_order=None):
    """The full render model for the matrix. PURE stat — no workbook opened (counts
    come from the caches). Per row, the SELECTED mode (`row_modes`, default 'env')
    decides each cell's comparison; `hidden`/`hidden_envs` drop rows/columns from
    the rendered+refreshed grid (still listed in all_rows/all_envs for the
    toggles); `row_order`/`env_order` are the user's drag-to-reorder preferences
    applied to the VISIBLE rows/columns. `row_defs` is injectable for tests."""
    now = now if now is not None else time.time()
    all_defs = row_defs if row_defs is not None else _row_defs()
    hidden = set(hidden or [])
    hidden_envs = set(hidden_envs or [])
    row_modes = row_modes or {}
    tsn_files = tsn_files or {}
    rows = {k: v for k, v in all_defs.items() if k not in hidden}
    rows = {k: rows[k] for k in apply_order(list(rows.keys()), row_order)}
    all_envs = list(envs) if envs is not None else env_keys()
    envs = apply_order([e for e in all_envs if e not in hidden_envs], env_order)
    env_labels = env_labels or {k: default_env_label(k) for k in all_envs}

    # Resolve each visible row's selected mode + gather the store subdirs whose
    # freshness the cells need (the mode's env side + any 'other' side).
    sel, needed = {}, set()
    for row_key, (label, subdir, _idx, adapter, _hr) in rows.items():
        mode = _mode_by_id(_row_modes(row_key, subdir, adapter),
                           row_modes.get(row_key, "env"))
        sel[row_key] = mode
        needed.add(mode.get("env_subdir", subdir))
        if mode.get("other_subdir"):
            needed.add(mode["other_subdir"])
    ages = report_library.cell_ages(dest, [(sd, sd) for sd in needed], envs, now=now)
    results = load_results(dest, baseline_key)
    tsn_results = load_tsn_results(dest)
    _absent = {"present": False, "mtime": None, "age_seconds": None}

    cells, modes_sel, modes_avail, tsn_meta = {}, {}, {}, {}
    for row_key, (label, subdir, _idx, adapter, _hr) in rows.items():
        mode = sel[row_key]
        modes_sel[row_key] = mode["id"]
        modes_avail[row_key] = [{"id": m["id"], "label": m["label"],
                                 "kind": m["kind"], "supported": m["supported"]}
                                for m in _row_modes(row_key, subdir, adapter)]
        env_subdir = mode.get("env_subdir", subdir)
        src = None
        if mode["kind"] == "tsn":
            src = (tsn_source(dest, mode["tsn_subdir"], tsn_files.get(mode["tsn_subdir"]))
                   if mode["supported"] else {"kind": "none"})
            tsn_meta[row_key] = {"supported": mode["supported"], "fmt": mode.get("fmt"),
                                 "source_kind": src.get("kind"), "source_path": src.get("path"),
                                 "pdf_count": src.get("pdf_count"),
                                 "tsn_subdir": mode["tsn_subdir"],
                                 "file": tsn_files.get(mode["tsn_subdir"]),
                                 "input_dir": str(tsn_input_root(dest, mode["tsn_subdir"]))}
        per = {}
        for env in envs:
            export = ages.get(env, {}).get(env_subdir, _absent)
            is_baseline = (mode["kind"] == "env" and env == baseline_key)
            if not mode["supported"]:
                cmp = {"supported": False}
            elif is_baseline:
                cmp = None
            elif mode["kind"] == "env":
                cmp = comparison_state(dest, baseline_key, row_key, env, env_subdir,
                                       ages, results)
            elif mode["kind"] == "tsn":
                rec = (tsn_results.get(f"{row_key}|{mode['id']}", {}) or {}).get(env)
                sources = [{"name": "cell", "present": export["present"], "mtime": export["mtime"]},
                           {"name": "tsn",
                            "present": src.get("kind") in ("file", "consolidated"),
                            "mtime": src.get("mtime")}]
                cmp = _cmp_state(mode_out_path(dest, baseline_key, row_key, env, mode),
                                 sources, rec)
            else:                            # self: TSMIS PDF vs Excel
                other = ages.get(env, {}).get(mode["other_subdir"], _absent)
                rec = (tsn_results.get(f"{row_key}|{mode['id']}", {}) or {}).get(env)
                sources = [{"name": "cell", "present": export["present"], "mtime": export["mtime"]},
                           {"name": "other", "present": other["present"], "mtime": other["mtime"]}]
                cmp = _cmp_state(mode_out_path(dest, baseline_key, row_key, env, mode),
                                 sources, rec)
            cell = {"export": export, "is_baseline": is_baseline, "cmp": cmp}
            if mode["kind"] == "env":
                cell["comparison"] = cmp     # back-compat alias for the Stage-A UI
            per[env] = cell
        cells[row_key] = per

    return {
        "dest": str(dest),
        "baseline": baseline_key,
        "rows": list(rows.keys()),
        "row_labels": {k: rows[k][0] for k in rows},
        # Every matrix row (visible or not) for the toggle chips; `tsn_capable`
        # marks the rows with a coded comparison beyond cross-env.
        "all_rows": [{"key": k, "label": all_defs[k][0],
                      "tsn_capable": tsn_capable(k)} for k in all_defs],
        "hidden": sorted(hidden),
        "modes": modes_sel,            # row_key -> selected mode id
        "row_modes": modes_avail,      # row_key -> [{id,label,kind,supported}]
        "tsn_meta": tsn_meta,          # row_key -> TSN source summary (tsn mode only)
        "envs": list(envs),
        "all_envs": all_envs,
        "hidden_envs": sorted(hidden_envs),
        "env_labels": env_labels,
        "cells": cells,
    }


# --------------------------------------------------------------------------- #
# optional live-formulas twin (opt-in). The matrix's offline counts + freshness
# all key off the VALUES workbook at the canonical out_path (compare_core is
# regression-locked, so we can't make mode="both" put values there). So when the
# user opts in, we ALSO write a recalculating formulas copy to a "(formulas)"
# sibling via a second compare pass — best-effort, never failing the values cell.
# --------------------------------------------------------------------------- #
def _formulas_sibling(out_path):
    out_path = Path(out_path)
    return out_path.with_name(f"{out_path.stem} (formulas){out_path.suffix}")


def _try_formulas(compare_call, out_path):
    """Run `compare_call(formulas_path)` (mode='formulas') beside the values copy.
    Best-effort: a failure here must NOT fail the already-written values cell."""
    try:
        compare_call(_formulas_sibling(out_path))
    except Exception as e:                       # noqa: BLE001
        log.warning("matrix: live-formulas workbook for %s not written (%s: %s)",
                    Path(out_path).name, type(e).__name__, e)


# --------------------------------------------------------------------------- #
# orchestration: build one cell's comparison (pure delegation to compare_env)
# --------------------------------------------------------------------------- #
def build_cell_comparison(dest, baseline_key, row_key, cell_key, events,
                          confirm_overwrite=None, row_defs=None, also_formulas=False):
    """Compare (row's report, cell env) against the baseline env, writing the
    VALUES workbook to comparison_path(...), and record its verdict + discrepancy
    counts in the cache. Pure delegation to the adapter's compare_folders — the
    comparison engine is untouched. Returns the ConsolidateResult. With
    `also_formulas`, also writes a live-formulas twin beside the values copy.

    Raises ValueError on an unknown row_key or a baseline cell (nothing to
    compare); compare_folders itself returns a clean error result when a side
    hasn't been exported yet."""
    rows = row_defs if row_defs is not None else _row_defs()
    if row_key not in rows:
        raise ValueError(f"unknown matrix row: {row_key}")
    if cell_key == baseline_key:
        raise ValueError("the baseline column has nothing to compare against")
    _label, subdir, _idx, adapter, has_route = rows[row_key]

    dest = Path(dest)
    out_path = comparison_path(dest, baseline_key, row_key, cell_key)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # side A = the cell env, side B = the baseline (labels read "ARS-TEST vs SSOR-PROD").
    result = adapter.compare_folders(
        dest / cell_key, dest / baseline_key, out_path,
        events=events, confirm_overwrite=confirm_overwrite or (lambda _p: True),
        mode="values")
    if also_formulas and result.status == "ok":
        _try_formulas(lambda fp: adapter.compare_folders(
            dest / cell_key, dest / baseline_key, fp, events=events,
            confirm_overwrite=lambda _p: True, mode="formulas"), out_path)

    if result.status == "ok" and out_path.exists():
        diff_cells, one_sided = read_counts(out_path, has_route)
        try:
            built_at = out_path.stat().st_mtime
        except OSError:
            built_at = None
        record_result(dest, baseline_key, row_key, cell_key, result.verdict,
                      diff_cells, one_sided, built_at)
    return result


def cells_to_rebuild(snapshot, scope="stale", row=None, env=None):
    """[(row_key, cell_key, mode_id)] to (re)compute, honoring each row's SELECTED
    mode. scope='all' = every comparable cell (both sides present); 'stale' = only
    missing/stale ones. Optional `row` / `env` filters drive the per-row and
    per-column refresh buttons. Skips the env-mode baseline column, unsupported
    (greyed) modes, and cells with a missing input side."""
    modes = snapshot.get("modes", {})
    todo = []
    for row_key in snapshot["rows"]:
        if row and row_key != row:
            continue
        mode_id = modes.get(row_key, "env")
        for ev in snapshot["envs"]:
            if env and ev != env:
                continue
            cmp = snapshot["cells"][row_key][ev].get("cmp")
            if cmp is None:                      # env-mode baseline column
                continue
            if not cmp.get("supported") or cmp.get("missing_side"):
                continue
            if scope == "all" or cmp.get("stale"):
                todo.append((row_key, ev, mode_id))
    return todo


# --------------------------------------------------------------------------- #
# orchestration: consolidate the env's store folder(s) and diff via the existing
# (untouched) comparison adapters. compare_highway_log / compare_highway_log_pdf
# are file-vs-file, so the per-route env folders are consolidated first.
# --------------------------------------------------------------------------- #
def _consolidate_store_folder(subdir, env_dir, out_path, events):
    """Consolidate one Export-Everything store folder (<env>/<subdir>/, per-route
    files) into a single workbook via the report's existing consolidator (with its
    additive input_dir/out_path override). Registry-driven via
    reports.consolidator_for_subdir, so any consolidatable report works; the PDF
    Highway Log is the one special case (needs a scratch converted_dir)."""
    out_path = Path(out_path)
    if subdir == "highway_log_pdf":
        import consolidate_tsmis_highway_log_pdf as _m   # pdfplumber — lazy
        conv = out_path.parent / f".{out_path.stem}_conv"
        try:
            _m.consolidate(events=events, confirm_overwrite=lambda _p: True,
                           input_dir=Path(env_dir), out_path=out_path,
                           converted_dir=conv)
        finally:
            shutil.rmtree(conv, ignore_errors=True)
        return
    import reports                                   # lazy (avoid import cycle)
    mod = reports.consolidator_for_subdir(subdir)
    if mod is None:
        raise ValueError(f"no store consolidator for {subdir}")
    mod.consolidate(events=events, confirm_overwrite=lambda _p: True,
                    input_dir=Path(env_dir), out_path=out_path)


def consolidate_tsn_pdfs(dest, subdir, events=None, confirm_overwrite=None):
    """Build a consolidated TSN workbook FROM the district PDFs the user dropped in
    <dest>/_tsn_input/<subdir>/, writing it back there so the next tsn_source()
    finds it. TSN is the SAME district-PDF set for both Highway Log flavors, so
    only the 'highway_log' TSN drop folder is handled. Returns the out path."""
    if subdir != "highway_log":
        raise ValueError(f"no TSN PDF consolidator for {subdir}")
    import consolidate_tsn_highway_log as _ctsn   # pdfplumber — lazy
    in_dir = tsn_input_root(dest, subdir)
    out_path = in_dir / "tsn_highway_log_consolidated.xlsx"
    _ctsn.consolidate(events=events, confirm_overwrite=confirm_overwrite or (lambda _p: True),
                      input_dir=in_dir, out_path=out_path)
    return out_path


# --- persistent (reusable) consolidated workbooks ------------------------- #
# v0.16.x: instead of consolidating a per-route store folder to a throwaway temp
# on EVERY comparison, persist the consolidated workbook into the run/store
# folder's `consolidated/` dir — the SAME filename + location the Consolidate tab
# uses — and REUSE it until the per-route files change. So: re-exporting a report
# makes its consolidated stale (a source file is newer) → the next comparison
# re-consolidates; only changing the comparison mechanism (not the data) reuses
# the existing consolidated. A force flag rebuilds it on demand.
def _consolidated_filename(subdir):
    if subdir == "highway_log_pdf":
        import consolidate_tsmis_highway_log_pdf as _m
        return _m.FILENAME
    import reports                                   # lazy
    mod = reports.consolidator_for_subdir(subdir)
    if mod is None:
        raise ValueError(f"no consolidated filename for {subdir}")
    return mod.FILENAME


def consolidated_store_path(store_dir, subdir):
    """The PERSISTENT consolidated workbook for a per-route store folder: a sibling
    `consolidated/` dir, date/env-stamped via the parent run-folder name (so a day
    folder gets a stamped name and matches the Consolidate-tab output; the always-
    current Everything store, whose parent is just `<src-env>`, gets the plain
    name)."""
    from paths import stamped_consolidated_filename     # lazy (avoid import cycle)
    parent = Path(store_dir).parent
    name = stamped_consolidated_filename(_consolidated_filename(subdir), parent.name)
    return parent / "consolidated" / name


def consolidated_state(store_dir, subdir):
    """{exists, fresh, path} for a store folder's persistent consolidated — drives
    the 'consolidated for this day' indicator. fresh = present AND newer than every
    per-route source file."""
    p = consolidated_store_path(store_dir, subdir)
    return {"exists": p.exists(), "fresh": not _consolidated_stale(p, store_dir),
            "path": str(p)}


def _consolidated_stale(consolidated, store_dir):
    """True when the persistent consolidated is missing, unreadable, or older than
    any per-route source file in the store folder (so it must be rebuilt)."""
    cm = _safe_mtime(consolidated)
    if cm is None:
        return True
    newest = None
    try:
        for e in Path(store_dir).iterdir():
            if e.is_file() and not e.name.startswith("~$"):
                m = e.stat().st_mtime
                if newest is None or m > newest:
                    newest = m
    except OSError:
        return True                                    # can't read store -> rebuild
    return newest is None or newest > cm + _MTIME_TOL_S


def consolidate_and_compare_tsn(tsmis_store_dir, tsn_path, out_path, row_key, subdir,
                                events, confirm_overwrite=None, force_consolidate=False,
                                also_formulas=False):
    """The SHARED TSN compare path used by BOTH matrices (the Everything matrix's
    latest-store cells AND the Compare tab's by-day cells).

    Consolidate a per-route TSMIS store folder (`tsmis_store_dir`, the `subdir`
    report) into its PERSISTENT `consolidated/` workbook (reused when still fresh;
    rebuilt when a source file is newer or `force_consolidate`), then compare it vs
    the consolidated TSN workbook with the row's comparator (`tsn_comparator_for(
    row_key)` — Highway Log Excel/PDF, Ramp Detail, …), writing the VALUES workbook
    to `out_path`. Returns the ConsolidateResult. Pure delegation to the untouched
    consolidate_* / compare_* adapters; the matrices differ only by the source
    folder + the output path. (For HL the output is byte-identical to the prior
    fmt-keyed path — same consolidator + same comparator.)"""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    cmp_mod = tsn_comparator_for(row_key)
    if cmp_mod is None:
        raise ValueError(f"no TSN comparator for {row_key}")
    consolidated = consolidated_store_path(tsmis_store_dir, subdir)
    if force_consolidate or _consolidated_stale(consolidated, tsmis_store_dir):
        _consolidate_store_folder(subdir, Path(tsmis_store_dir), consolidated, events)
    # The consolidator can return without raising yet write nothing (an empty store
    # folder). Catch that here so the failure names the consolidation step instead
    # of the compare adapter raising a confusing error on a missing/empty input.
    if not consolidated.exists() or consolidated.stat().st_size == 0:
        raise ValueError(f"nothing to compare — no {subdir} export found in "
                         f"{tsmis_store_dir}")
    result = cmp_mod.compare(consolidated, str(tsn_path), out_path, events=events,
                             confirm_overwrite=confirm_overwrite or (lambda _p: True),
                             mode="values")
    if also_formulas and result.status == "ok":
        _try_formulas(lambda fp: cmp_mod.compare(
            consolidated, str(tsn_path), fp, events=events,
            confirm_overwrite=lambda _p: True, mode="formulas"), out_path)
    return result


def _ensure_consolidated(store_dir, subdir, events, force):
    """Return the persistent consolidated workbook for a store folder, rebuilding
    it when stale or forced. Shared by the self (PDF-vs-Excel) path."""
    p = consolidated_store_path(store_dir, subdir)
    if force or _consolidated_stale(p, store_dir):
        _consolidate_store_folder(subdir, Path(store_dir), p, events)
    return p


def build_comparison(dest, row_key, cell_key, mode_id, baseline_key, events,
                     tsn_files=None, confirm_overwrite=None, row_defs=None,
                     force_consolidate=False, also_formulas=False):
    """(Re)build one cell's comparison for the row's SELECTED mode, write the VALUES
    workbook, and cache its counts. Dispatches to the existing comparison adapters
    (never edits them). Returns the ConsolidateResult. Raises ValueError for an
    unknown row or an unsupported/greyed mode; an absent input side yields the
    adapter's clean error result. `force_consolidate` rebuilds the persistent
    consolidated even when it looks fresh; `also_formulas` writes a live-formulas
    twin beside the values copy."""
    rows = row_defs if row_defs is not None else _row_defs()
    if row_key not in rows:
        raise ValueError(f"unknown matrix row: {row_key}")
    _label, subdir, _idx, adapter, _hr = rows[row_key]
    mode = _mode_by_id(_row_modes(row_key, subdir, adapter), mode_id)
    if not mode["supported"]:
        raise ValueError(f"no comparison for {row_key} / {mode['id']}")

    if mode["kind"] == "env":
        return build_cell_comparison(dest, baseline_key, row_key, cell_key, events,
                                     confirm_overwrite=confirm_overwrite, row_defs=rows,
                                     also_formulas=also_formulas)

    dest = Path(dest)
    out_path = mode_out_path(dest, baseline_key, row_key, cell_key, mode)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tsn_files = tsn_files or {}
    if mode["kind"] == "tsn":
        src = tsn_source(dest, mode["tsn_subdir"], tsn_files.get(mode["tsn_subdir"]))
        if src.get("kind") not in ("file", "consolidated"):
            raise ValueError("no consolidated TSN workbook available")
        result = consolidate_and_compare_tsn(
            dest / cell_key / mode["env_subdir"], src["path"], out_path,
            row_key, mode["env_subdir"], events, confirm_overwrite=confirm_overwrite,
            force_consolidate=force_consolidate, also_formulas=also_formulas)
    else:                                # self: TSMIS PDF vs Excel (persisted both sides)
        side_env = _ensure_consolidated(dest / cell_key / mode["env_subdir"],
                                        mode["env_subdir"], events, force_consolidate)
        side_other = _ensure_consolidated(dest / cell_key / mode["other_subdir"],
                                          mode["other_subdir"], events, force_consolidate)
        # the adapter fixes PDF=side A, Excel=side B regardless of the row.
        pdf_c = side_env if mode["env_subdir"] == "highway_log_pdf" else side_other
        excel_c = side_other if mode["env_subdir"] == "highway_log_pdf" else side_env
        import compare_highway_log_pdf as _cmp_p
        result = _cmp_p.TSMIS_PDF_VS_EXCEL.compare(
            pdf_c, excel_c, out_path, events=events,
            confirm_overwrite=confirm_overwrite or (lambda _p: True), mode="values")
        if also_formulas and result.status == "ok":
            _try_formulas(lambda fp: _cmp_p.TSMIS_PDF_VS_EXCEL.compare(
                pdf_c, excel_c, fp, events=events,
                confirm_overwrite=lambda _p: True, mode="formulas"), out_path)

    if result.status == "ok" and out_path.exists():
        # has_route=True is correct for ALL tsn/self modes: their output is always
        # the consolidated Highway Log shape (Route-keyed), regardless of the row's
        # cross-env `_hr` (the HL-PDF row has adapter=None → _hr=False, but its TSN
        # output is still Route-keyed). Do NOT switch this to the row's _hr.
        diff_cells, one_sided = read_counts(out_path, has_route=True)
        record_tsn_result(dest, f"{row_key}|{mode['id']}", cell_key, result.verdict,
                          diff_cells, one_sided, _safe_mtime(out_path))
    return result
