"""Highway Detail adapter for the visual-evidence generator (visual_evidence).

Supplies everything report-specific the engine needs to turn a vs-TSN diff into
a pair of highlighted PDF snippets:

  * the DIFF SOURCE — both comparison sides re-loaded through the comparator's
    OWN loaders/normalizations (compare_highway_detail_tsn), so an "example" is
    exactly a cell the comparison flags, never a re-derivation that could drift;
  * the TSMIS locator — the same per-page window walk the Highway Detail PDF
    consolidator uses, kept in LOCKSTEP with it (see _locate note), but keeping
    each record's page / y-extent / column geometry so a field maps to a cell
    rectangle;
  * the TSN locator — the district TASAS print parsed line-anchored (the same
    two-line record regexes validated against the statewide extract, >=99.9%
    field agreement), with word positions kept so a regex group maps to a box;
  * VERIFICATION projections — a candidate is only usable when the value parsed
    back out of each PDF, run through the comparator's own projections, equals
    the value the comparison compared (so an evidence image can never show
    something other than what was diffed; the known render skews are skipped
    with a logged reason instead).

Row identity matches the comparison: the canonical roadbed-aware Post Mile key,
restricted to keys UNIQUE on both sides of a route so a highlight is THE row
(the comparison's occurrence pairing handles duplicates; evidence just avoids
them). Console-free; pdfplumber/openpyxl gated by the engine.
"""
import logging
import re
from collections import defaultdict
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import load_workbook
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import compare_highway_detail_tsn as cht
import consolidate_tsmis_highway_detail_pdf as chd
import highway_detail_columns as hdc
from pdf_table_lib import (cluster_by_top, median, norm_route,
                           require_document_route)
from tsn_load_highway_detail import SIDECAR_HEADER, tsn_rows_with_dcr  # noqa: F401

log = logging.getLogger("tsmis.evidence")

REPORT_LABEL = "Highway Detail"
KEY_LABEL = cht.KEY
# Every compared column except the key itself (PS included — the equation
# marker the two systems print in different places is evidence gold).
FIELDS = [f for f in cht.SHARED_HEADER if f != cht.KEY]

_S = cht._s  # the comparator's None-safe strip — evidence must normalize alike


# --------------------------------------------------------------------------- #
# TSN district print layout (the legacy TASAS two-line record)
# --------------------------------------------------------------------------- #
DCR_RE = re.compile(r"DIST-CNTY-ROUTE\s+(\d{2})-([A-Z]{1,3})-(\d{3})(?:\s+([A-Z]))?")
# TSN DITTO marks (the Highway Log convention, in Highway Detail too): a fully
# continued roadbed/median block prints WIDTH-MATCHED runs of '+' — a dittoed
# effective date is '++++++++' (8 chars), a lane count '++', a 3-digit width
# '+++' — and the statewide extract stores the same '+'-run TEXT, so the
# comparison flags those cells as ordinary text diffs. The locator must parse a
# ditto record (real D04 print: "… 30V ++++++++ + ++ + ++ ++ +++ ++ ++"), or
# every such row becomes unlocatable evidence.
_DATE = r"(?:\d{2}-\d{2}-\d{2}|\+{1,8})"
_NUM = r"(?:\d{1,3}|\+{1,3})"
# line 1: PP? MILE PS? LENGTH REC_DATE HG AC [sig]ACC_EFF [CITY] RU BEG_DATE <ADT tail>
L1_RE = re.compile(
    rf"^\s*(?P<pp>[A-Z]\s+)?(?P<mile>\d{{3}}\.\d{{3}})(?P<ps>[A-Z])?\s+"
    rf"(?P<len>\d{{3}}\.\d{{3}})\s+(?P<rec>{_DATE})\s+(?P<hg>[A-Z])\s+(?P<ac>[A-Z])\s+"
    rf"(?P<accsig>[*Y])?(?P<aceff>{_DATE})\s+(?:(?P<city>[A-Z][A-Z0-9]{{0,3}})\s+)?"
    rf"(?P<ru>[RUB])\s+(?P<beg>{_DATE})\s+(?P<adt>\S+)(?P<tail>.*)$")
# line 2, left-lazy: DESC? NA [sig]LB-EFF ST LN SF OT-TO OT-TR WID IN-TO IN-TR
#                    [sig]MED-EFF T C B WDA [sig]RB-EFF ST LN SF IN-TO IN-TR WID OT-TO OT-TR
# (the RB half is MIRRORED around the median in the print — inner before width
# before outer — which is also the TSMIS export's column order).
L2_RE = re.compile(
    rf"^(?P<desc>.*?)\s*(?P<na>[AN])\s+"
    rf"(?P<lbsig>[*Y])?(?P<lbeff>{_DATE})\s+(?P<lbt>\S)\s+(?P<lbln>{_NUM})\s+(?P<lbf>\S)\s+"
    rf"(?P<lbto1>{_NUM})\s+(?P<lbtr1>{_NUM})\s+(?P<lbwid>{_NUM})\s+(?P<lbto2>{_NUM})\s+(?P<lbtr2>{_NUM})\s+"
    rf"(?P<medsig>[*Y])?(?P<medeff>{_DATE})\s+(?P<medt>\S)\s+(?P<medc>\S)\s+(?P<medb>\S)\s+"
    rf"(?P<medwda>(?:\d{{1,3}}[A-Z]?|\+{{1,4}}))\s+"
    rf"(?P<rbsig>[*Y])?(?P<rbeff>{_DATE})\s+(?P<rbt>\S)\s+(?P<rbln>{_NUM})\s+(?P<rbf>\S)\s+"
    rf"(?P<rbto1>{_NUM})\s+(?P<rbtr1>{_NUM})\s+(?P<rbwid>{_NUM})\s+(?P<rbto2>{_NUM})\s+(?P<rbtr2>{_NUM})\s*$")

# shared comparison field -> TSN regex group.
TSN_GROUP = {
    "PS": "ps", "Length": "len", "Date of Rec": "rec", "HG": "hg", "AC": "ac",
    "Acc-Cont Eff": "aceff", "City": "city", "RU": "ru", "RU Eff": "beg",
    "Description": "desc", "NA": "na",
    "LB Eff": "lbeff", "LB S/T": "lbt", "LB #Ln": "lbln", "LB S/F": "lbf",
    "LB OT-TO": "lbto1", "LB OT-TR": "lbtr1", "LB Wid": "lbwid",
    "LB IN-TO": "lbto2", "LB IN-TR": "lbtr2",
    "Med Eff": "medeff", "Med T": "medt", "Med C": "medc", "Med B": "medb",
    "Med V/WDA": "medwda",
    "RB Eff": "rbeff", "RB S/T": "rbt", "RB #Ln": "rbln", "RB S/F": "rbf",
    "RB IN-TO": "rbto1", "RB IN-TR": "rbtr1", "RB Wid": "rbwid",
    "RB OT-TO": "rbto2", "RB OT-TR": "rbtr2",
}
_L1_GROUPS = {"ps", "len", "rec", "hg", "ac", "aceff", "city", "ru", "beg"}
_L1_ORDER = ["pp", "mile", "ps", "len", "rec", "hg", "ac", "accsig", "aceff",
             "city", "ru", "beg", "adt"]
_L2_ORDER = ["desc", "na", "lbsig", "lbeff", "lbt", "lbln", "lbf", "lbto1",
             "lbtr1", "lbwid", "lbto2", "lbtr2", "medsig", "medeff", "medt",
             "medc", "medb", "medwda", "rbsig", "rbeff", "rbt", "rbln", "rbf",
             "rbto1", "rbtr1", "rbwid", "rbto2", "rbtr2"]

# --------------------------------------------------------------------------- #
# diff source — both sides through the comparator's own loaders
# --------------------------------------------------------------------------- #
def load_sides(consolidated_path, tsn_path):
    """(tsmis_rows, tsn_rows, sidecar, note) — both sides in the comparator's
    shape ([route, *SHARED_HEADER]). `sidecar` maps (route, key) -> [(district,
    county), ...] for locating TSN rows in the district prints; None when the
    TSN workbook carries no district info (an old normalized library), with
    `note` saying what to do about it."""
    tsmis_rows, _route_keyed = cht._load_tsmis(consolidated_path)
    tsn_rows, sidecar, note = _load_tsn_with_sidecar(tsn_path)
    return tsmis_rows, tsn_rows, sidecar, note


def _load_tsn_with_sidecar(path):
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if cht.NORMALIZED_SHEET in wb.sheetnames:
            it = wb[cht.NORMALIZED_SHEET].iter_rows(values_only=True)
            header = list(next(it, []) or [])
            has_sidecar = len(header) >= 1 + len(cht.SHARED_HEADER) + 2
            rows, sidecar = [], defaultdict(list)
            for r in it:
                if not r or all(c in (None, "") for c in r):
                    continue
                row = cht._normalized_row(r)   # re-projected like the comparison
                rows.append(row)
                if has_sidecar:
                    dist = _S(r[1 + len(cht.SHARED_HEADER)])
                    cnty = _S(r[2 + len(cht.SHARED_HEADER)]).rstrip(".")
                    sidecar[(row[0], row[1])].append((dist, cnty))
            if not has_sidecar:
                return rows, None, ("the normalized TSN library predates the "
                                    "evidence columns — rebuild the TSN library "
                                    "(Settings) and run the comparison again")
            return rows, sidecar, None
    finally:
        wb.close()
    rows, dcr = tsn_rows_with_dcr(path)          # a RAW statewide file
    sidecar = defaultdict(list)
    for row, (dist, cnty) in zip(rows, dcr):
        sidecar[(row[0], row[1])].append((dist, cnty))
    return rows, sidecar, None


def enumerate_diffs(tsmis_rows, tsn_rows, sidecar):
    """{field: [example]} over keys UNIQUE per route on BOTH sides — each
    example a cell the comparison flags (values already normalized by the
    loaders, so inequality here == a red cell there)."""
    a_route, b_route = defaultdict(list), defaultdict(list)
    for r in tsmis_rows:
        a_route[r[0]].append(r)
    for r in tsn_rows:
        b_route[r[0]].append(r)
    diffs = defaultdict(list)
    for route in sorted(set(a_route) & set(b_route)):
        a_ct, b_ct = defaultdict(int), defaultdict(int)
        for r in a_route[route]:
            a_ct[r[1]] += 1
        for r in b_route[route]:
            b_ct[r[1]] += 1
        a_by = {r[1]: r for r in a_route[route] if a_ct[r[1]] == 1}
        b_by = {r[1]: r for r in b_route[route] if b_ct[r[1]] == 1}
        for key in set(a_by) & set(b_by):
            ra, rb = a_by[key], b_by[key]
            for i, f in enumerate(cht.SHARED_HEADER):
                if f == cht.KEY:
                    continue
                va, vb = _S(ra[1 + i]), _S(rb[1 + i])
                if va != vb:
                    dist, cnty = (sidecar.get((route, key)) or [("", "")])[0]
                    diffs[f].append(dict(route=route, key=key, field=f,
                                         va=va, vb=vb, dist=dist, cnty=cnty))
    return diffs


# --------------------------------------------------------------------------- #
# verification projections
# --------------------------------------------------------------------------- #
def project(field, raw):
    """A raw PDF cell value -> the compared form, via the comparator's own
    per-field projection (PS is key-derived, not a plain column)."""
    if field == "PS":
        return "E" if "E" in _S(raw).upper() else ""
    return cht._project(field, raw)


# --------------------------------------------------------------------------- #
# TSMIS side — the per-route "Highway Detail (PDF)" export
# --------------------------------------------------------------------------- #
def tsmis_pdf_path(pdf_dir, route):
    return Path(pdf_dir) / f"highway_detail_route_{route}.pdf"


def _page_edges(page):
    """Median (x0,x1) CELL edges per column for both band shapes — the box for
    a BLANK cell, where no character can supply a bbox."""
    bands = chd._page_bands(page)

    def edges(bl, n):
        if not bl:
            return None
        return list(zip([median([b[i]["x0"] for b in bl]) for i in range(n)],
                        [median([b[i]["x1"] for b in bl]) for i in range(n)]))

    return (edges(bands.get(chd.N_COLS_L1), chd.N_COLS_L1),
            edges(bands.get(chd.N_COLS_L2), chd.N_COLS_L2))


def _gmeta(group, page_no, win, edges, fb):
    chars = [c for _t, cs in group for c in cs]
    return {"page": page_no, "win": win, "edges": edges, "fb": fb,
            "chars": chars,
            "top": min(c["top"] for c in chars),
            "bottom": max(c["bottom"] for c in chars)}


def locate_tsmis(pdf_path, needed_keys):
    """{canonical_key: [record]} for `needed_keys`; a record carries the parsed
    34-column row plus per-line meta (page, y-extent, windows, edges, chars).

    LOCKSTEP: this walk mirrors consolidate_tsmis_highway_detail_pdf.parse_pdf
    step for step (per-page windows, physical row groups, the postmile line-1
    test, the DATE_TOKEN furniture guard, the cross-page pending carry, the
    document-median fallback, the CMP-AUD-049 banner-claim capture BEFORE the
    geometry gate) — it only ADDS position capture. A behavior change there
    must land here too; check_visual_evidence pins the shared pieces so a
    drift fails the gate.

    CMP-AUD-049 (evidence half): raises pdf_table_lib.RouteIdentityError when
    the document's own page-banner claims don't confirm the route the
    filename names."""
    found = defaultdict(list)
    doc_routes = set()                 # the pages' own route claims (049)
    fm = re.search(r"route_([0-9A-Za-z]+)\.pdf$", str(pdf_path))
    file_route = fm.group(1) if fm else None
    doc_win = {}
    pending = pending_meta = None
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            groups = chd._row_groups(page)
            for group in groups:
                bm = chd.BANNER_ROUTE_RE.match(chd._group_text(group))
                if bm:
                    doc_routes.add(bm.group(1))
            win1, win2 = chd._page_windows(page)
            fb = False
            if win1 is None or win2 is None:
                if "w1" not in doc_win:
                    doc_win["w1"] = chd._document_windows(pdf, chd.N_COLS_L1)
                    doc_win["w2"] = chd._document_windows(pdf, chd.N_COLS_L2)
                win1, win2 = win1 or doc_win["w1"], win2 or doc_win["w2"]
                if win1 is None or win2 is None:
                    continue
                fb = True
            e1, e2 = _page_edges(page)
            for group in groups:
                vals1 = chd._group_values(group, win1)
                if chd._is_line1(vals1):
                    pending = vals1
                    pending_meta = _gmeta(group, page_no, win1, e1, fb)
                elif pending is not None:
                    vals2 = chd._group_values(group, win2)
                    if not any(v and chd.DATE_TOKEN_RE.search(v) for v in vals2):
                        continue
                    row = chd._make_row(pending, vals2)
                    canon = cht.pm_canon(row[0] or "", row[3] or "")
                    if canon in needed_keys:
                        found[canon].append(
                            {"row": row, "m1": pending_meta,
                             "m2": _gmeta(group, page_no, win2, e2, fb)})
                    pending = pending_meta = None
    require_document_route(
        Path(pdf_path).name, norm_route(file_route) if file_route else None,
        [norm_route(t) for t in doc_routes],
        claim_desc="the page banner's \"Ref Date: … Route NNN Page N\"")
    return found


def tsmis_value(rec, field):
    """The compared value this PDF record carries for `field` (verification)."""
    if field == "PS":
        return cht.pm_suffix(rec["row"][0] or "")
    return project(field, rec["row"][hdc.HEADER.index(field)])


def tsmis_box(rec, field):
    """(page_no, cell_box, record_yspan, table_xspan) for `field`'s cell.
    Rejects (returns None) a record parsed on the fallback grid or split
    across pages — the geometry there is approximate, not evidence-grade."""
    if rec["m1"]["fb"] or rec["m2"]["fb"]:
        return None
    if rec["m1"]["page"] != rec["m2"]["page"]:
        return None
    if field in (cht.KEY, "PS"):
        line, idx = 1, 0
    else:
        pos = hdc.HEADER.index(field)
        line, idx = (1, pos) if pos <= 8 else (2, pos - 9)
    meta = rec["m1"] if line == 1 else rec["m2"]
    lo, hi = meta["win"][idx]
    hits = [c for c in meta["chars"] if lo <= (c["x0"] + c["x1"]) / 2 < hi]
    if hits:
        x0, x1 = min(c["x0"] for c in hits), max(c["x1"] for c in hits)
    elif meta["edges"]:
        x0, x1 = meta["edges"][idx]
    else:
        return None
    e1, e2 = rec["m1"]["edges"], rec["m2"]["edges"]
    xs = [e[0][0] for e in (e1, e2) if e] + [e[-1][1] for e in (e1, e2) if e]
    if not xs:
        return None
    return (meta["page"],
            (x0 - 2, meta["top"] - 2, x1 + 2, meta["bottom"] + 2),
            (rec["m1"]["top"], rec["m2"]["bottom"]),
            (min(xs) - 4, max(xs) + 4))


# --------------------------------------------------------------------------- #
# TSN side — the district TASAS prints
# --------------------------------------------------------------------------- #
def district_index(pdf_dir, events=None):
    """{district('01'..'12'): path} by reading each PDF's own first
    DIST-CNTY-ROUTE header (filenames are the user's business). A PDF with no
    DCR header in its first pages is skipped with a log line."""
    index = {}
    for p in sorted(Path(pdf_dir).glob("*.pdf")):
        dist = None
        try:
            with pdfplumber.open(p) as pdf:
                for page in pdf.pages[:3]:
                    m = DCR_RE.search(page.extract_text() or "")
                    if m:
                        dist = m.group(1)
                        break
        except Exception as e:                            # unreadable/odd PDF
            log.warning("evidence: %s unreadable: %s: %s",
                        p.name, type(e).__name__, e)
            if events:
                events.on_log(f"    note: {p.name} unreadable, skipped")
            continue
        if dist is None:
            log.info("evidence: %s has no DIST-CNTY-ROUTE header; skipped", p.name)
            continue
        index.setdefault(dist, p)
    return index


def _word_lines(page):
    out = []
    for _top, ws in cluster_by_top(page.extract_words(), 3):
        parts, offs, pos = [], [], 0
        for w in ws:
            t = w["text"]
            offs.append((pos, pos + len(t), w))
            parts.append(t)
            pos += len(t) + 1
        out.append({"text": " ".join(parts), "offs": offs,
                    "top": min(w["top"] for w in ws),
                    "bottom": max(w["bottom"] for w in ws)})
    return out


def locate_tsn(pdf_path, needed_routes, needed_keys):
    """{(county, route, canonical_key): [record]} from one district print,
    scanning only lines under a DCR header whose route is needed. A record
    keeps both lines' regex matches, word offsets, pages and y-extents."""
    found = defaultdict(list)
    dcr = None
    pending = None
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            for ln in _word_lines(page):
                text = ln["text"]
                m = DCR_RE.search(text)
                if m:
                    dcr = (m.group(2), norm_route(m.group(3) + (m.group(4) or "")))
                    pending = None
                    continue
                if dcr is None or dcr[1] not in needed_routes:
                    continue
                m1 = L1_RE.match(text)
                if m1 and pending is None:
                    pending = (m1, ln, page_no)
                    continue
                if pending is not None:
                    m2 = L2_RE.match(text)
                    if m2:
                        d = pending[0].groupdict()
                        token = f"{_S(d['pp'])}{d['mile']}{_S(d['ps'])}"
                        key3 = (dcr[0], dcr[1], cht.pm_canon(token, d["hg"]))
                        if key3 in needed_keys:
                            found[key3].append(
                                {"m1": pending[0], "l1": pending[1],
                                 "p1": pending[2],
                                 "m2": m2, "l2": ln, "p2": page_no})
                        pending = None
                    else:
                        # furniture, or a new line 1 replacing a dangling one
                        m1b = L1_RE.match(text)
                        pending = (m1b, ln, page_no) if m1b else pending
    return found


def _span_box(ln, s, e):
    """x-range of text span [s,e) on a word-indexed line; an EMPTY span (an
    optional group that didn't print) boxes the gap between its neighbors."""
    ws = [w for a, b, w in ln["offs"] if a < e and b > s]
    if ws:
        return min(w["x0"] for w in ws), max(w["x1"] for w in ws)
    left = [w for a, b, w in ln["offs"] if b <= s]
    right = [w for a, b, w in ln["offs"] if a >= e]
    x0 = left[-1]["x1"] + 1 if left else (right[0]["x0"] - 12 if right else 0)
    x1 = right[0]["x0"] - 1 if right else x0 + 12
    return (x0, x1) if x1 > x0 else (x0, x0 + 10)


def tsn_value(rec, field):
    """The compared value this print record carries for `field`."""
    g = TSN_GROUP[field]
    m = rec["m1"] if g in _L1_GROUPS else rec["m2"]
    raw = m.group(g) if m.span(g)[0] != -1 else ""
    return project(field, raw)


def tsn_box(rec, field):
    """(page_no, cell_box, record_yspan, record_xspan) for `field`'s cell."""
    g = TSN_GROUP[field]
    is_l1 = g in _L1_GROUPS
    m, ln, page_no = ((rec["m1"], rec["l1"], rec["p1"]) if is_l1 else
                      (rec["m2"], rec["l2"], rec["p2"]))
    order = _L1_ORDER if is_l1 else _L2_ORDER
    s, e = m.span(g)
    if s == -1:                                   # optional group not printed
        i = order.index(g)
        prev = next((m.span(o)[1] for o in reversed(order[:i])
                     if m.span(o)[0] != -1), 0)
        nxt = next((m.span(o)[0] for o in order[i + 1:]
                    if m.span(o)[0] != -1), len(ln["text"]))
        s = e = min(prev + 1, nxt)
    x0, x1 = _span_box(ln, s, e)
    words = rec["l1"]["offs"] + rec["l2"]["offs"]
    xspan = (min(w["x0"] for _a, _b, w in words) - 4,
             max(w["x1"] for _a, _b, w in words) + 4)
    # a record split across pages keeps the target line's own extent
    yspan = ((rec["l1"]["top"], rec["l2"]["bottom"])
             if rec["p1"] == rec["p2"] else (ln["top"], ln["bottom"]))
    return page_no, (x0 - 2, ln["top"] - 2, x1 + 2, ln["bottom"] + 2), yspan, xspan
