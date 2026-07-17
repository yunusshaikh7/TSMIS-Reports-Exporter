"""Build the TSMIS-vs-TSN Intersection Detail discrepancy workbook (FLAT, route+PM).

The Ramp Detail FLAT recipe applied to Intersection Detail: both sides are XLSX in
different shapes, each with its own loader projecting onto ONE shared PM-keyed
header. Reconciled by hand on route 1, then re-reconciled statewide (16,200 paired
rows) against the July 2026 site update, which reshaped the export:

  * TSMIS side — the CONSOLIDATED Intersection Detail workbook (leading Route
    column + the 35 source columns of the July-2026 format). Some header labels
    still sit shifted against their values (the "INT Type" label is over the
    eff-date, etc.), so columns are read BY POSITION; PRE-update workbooks (the
    old 36-column layout with the duplicated second 'ML Eff-Date') are REFUSED
    with a re-export hint rather than mis-read.
  * TSN side — the statewide raw `Sheet 1` (36 named DB columns); route from
    `LOCATION` ("12 ORA 001" -> "001").

Both store attribute values in (eff_date, type) order (the planning-phase
"pair-order reversal" was a misread of the shifted TSMIS labels).

Every field present in both systems is COMPARED and COUNTED — a mechanical diff; the
reader adds commentary, the tool never hides a column. Columns are ordered to mirror the
source report and every column is compared BY REPORT POSITION — each report column to the
same column in the other report (user decision 2026-06-24); nothing is suppressed. The
July 2026 update fixed most of the old structural date classes (Date of Record and the
INT / Control / Lighting eff-dates now match TSN on ≥99.9% of rows — the old ~1-day
offset is gone; the booleans are natively Y/N; the postmiles print zero-padded; the
Location now carries the route suffix): what REMAINS structural is Int St Eff-Date
(TSN stores a bulk refresh stamp — '2022-01-01' on ~99% of rows — where TSMIS keeps the
historical date) and, to a lesser degree, ML / CS Eff-Date (~12% / ~3%: TSN tracks a
LATER resurvey date where TSMIS keeps the original). 'Xing Line Lgth' (TSN's
X_CROSS_OVERRIDE) is newly exported and newly compared; TSMIS no longer exports the
second ML Eff-Date, so TSN's MAIN_EFF_DATE is now a TSN-only reference column (blue on
the Report View, like the ADT pair). A second "Report View" sheet replicates the printed
two-line record and shows every difference in red; the three structural date columns and
the route suffix stay OUT of its per-record "Major" count (user decision 2026-07-08 —
the data-driven soft set). Normalizations make some raw-different values compare equal;
each is documented in the Notes sheet so a match is read as "equal after the stated
normalization", not raw equality:
  1. **Control-type crosswalk** — TSN records signalized under the legacy signal
     sub-types J/K/L/M/N/P, which TSNR/TSMIS collapses into the single category TSMIS
     stores as the code "S". Per the TSNR/MIRE reference, both sides' signalized codes
     are normalized to that one code "S" (the Signalized category) so the sub-type
     split stops flagging; the Notes sheet documents it. (Geometry/INT Type needs no
     crosswalk — both systems share the F/M/S/T/Y/Z/R codes.)
  2. **Boolean encoding** — both systems now store Y/N; PRE-update TSMIS data stored
     1/0, and the same Y≡1 / N≡0 normalization still applies so only genuine changes
     flag wherever legacy-encoded values appear.
  3. **Numeric zero-padding** — Main Line Length, the intersecting-route block and
     Xing Line Lgth canonicalize ('058'≡'58', '9.560'≡'9.56') so padding never flags.

Console-free; engine in compare_core.
"""
import dataclasses
import logging
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import CellRange
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

import compare_tsn_common as ctc
import comparison_contract as cc
from compare_tsn_common import (load_consolidated_rows, row_has_data,
                                suggest_route_name)
from compare_core import (CompareSchema, compared_cell, normalize_value, keys_for,
                          pair_occurrences_by_similarity, union_keys,
                          set_safe_literal_cell, _PROGRESS_EVERY)


log = logging.getLogger("tsmis.compare")
REPORT_NAME = "Intersection Detail"
TSMIS_SHEET = "Intersection Detail"      # consolidated sheet (Route prepended)
TSN_SHEET = "Sheet 1"                     # raw statewide DB dump
NORMALIZED_SHEET = "Intersection Detail (TSN)"
# CMP-AUD-037: the DIRECT-path freshness marker version (the catalog's
# intersection_detail normalization_version MIRRORS this; tsn_load_
# intersection_detail.build_into stamps it, _load_tsn refuses anything older).
# v5 is a marker-only bump over the v4 county-aware shape — the normalized rows
# are byte-identical, but a bare v4 library carried no marker and the direct
# path trusted it; the bump forces D2 to rebuild stored libraries so they gain it.
NORMALIZATION_VERSION = 5
# CMP-AUD-033: the documented sidecar columns that follow ["Route"] +
# SHARED_HEADER in the normalized workbook (tsn_load_intersection_detail
# .SIDECAR_HEADER mirrors this; check_tsn_normalization_marker gates the mirror).
_NORMALIZED_SIDECARS = ("TSN District", "TSN County")

KEY = "PM"
# Column order mirrors the source report (each effective-date next to its type, the
# mainline block, then the cross-street block, then the intersecting route). Every
# field present in both systems is compared and counted — nothing is suppressed
# (CONTEXT_FIELDS is empty, below; the position-aligned policy compares every column).
# District + County are COMPARED fields (the accepted ID-79 oracle asserts them
# — 34 asserting fields per paired row; both derive from each source's Location
# and are construction-equal on paired rows, so they add assertion coverage and
# visible provenance without changing any difference count).
SHARED_HEADER = [
    "PR", "Route Suffix", "PM", "District", "County",
    "Date of Record", "HG", "City Code", "R/U",
    "INT Type Eff-Date", "INT Type",
    "Control Type Eff-Date", "Control Type",
    "Lighting Eff-Date", "Lighting",
    "ML Eff-Date", "ML Mastarm", "ML Left Chan", "ML Right Chan",
    "ML Traffic Flow", "ML Num Lanes",
    "Description", "Main Line Length",
    "CS Eff-Date", "CS Mastarm", "CS Left Chan", "CS Right Chan",
    "CS Traffic Flow", "CS Num Lanes", "Int St Eff-Date",
    "Intrte Route", "Intrte PM Prefix", "Intrte Postmile", "Intrte PM Suffix",
    "Xing Line Lgth",
]
KEY_FIELD = SHARED_HEADER.index(KEY)      # 2 (after PR + the derived Route Suffix column)
# Position-aligned comparison (user decision 2026-06-24): every report column is compared
# to the same column in the other report — nothing is suppressed. Since the July 2026
# export update only Int St Eff-Date differs structurally (TSN's bulk refresh stamp vs
# TSMIS's historical date) plus a smaller ML/CS eff-date resurvey-tracking gap; the Notes
# sheet documents them. TSMIS no longer exports the second ML Eff-Date, so TSN's
# MAIN_EFF_DATE is a Report-View-only reference column now (with the ADT pair).
CONTEXT_FIELDS = ()
DATE_FIELDS = ("Date of Record", "INT Type Eff-Date", "Control Type Eff-Date",
               "Lighting Eff-Date", "ML Eff-Date", "CS Eff-Date",
               "Int St Eff-Date")
# Both systems store Y/N since the July 2026 update; pre-update TSMIS data stored 1/0
# and the same normalization still folds it, so only real changes flag either way.
BOOLEAN_FIELDS = ("Lighting", "ML Mastarm", "ML Right Chan", "CS Mastarm", "CS Right Chan")
# Numeric fields where the two systems differ only in zero-padding — Main Line Length
# (TSMIS '58' vs TSN '058'), the intersecting-route number, its postmile (TSMIS
# '9.560' vs TSN '9.56'), and the crossing line length (X_CROSS_OVERRIDE).
# Normalized to a canonical number so the padding doesn't flag.
NUMERIC_FIELDS = ("Main Line Length", "Intrte Route", "Intrte Postmile", "Xing Line Lgth")
# Control-type crosswalk (per the TSNR/MIRE reference "TSNR - Intersection Control
# and Geometry Type"): TSN spreads the signalized category across the legacy signal
# sub-types J–P (J/K/L/M/N/P), which TSNR/TSMIS collapses into the single code TSMIS
# stores as "S". Both sides' signalized codes normalize to that one code "S" (the
# Signalized category) — so the compared Control Type cell shows "S" wherever the
# crosswalk applied, and the sub-type split stops flagging as a difference. The Notes
# sheet documents the crosswalk. Geometry (INT Type) needs NO crosswalk — both systems
# share F/M/S/T/Y/Z/R.
_SIGNALIZED_CODES = {"J", "K", "L", "M", "N", "P", "S"}
_SIGNALIZED_LABEL = "S"          # TSN J–P + TSMIS S all fold to TSMIS's code "S"

# TSN raw column name for each shared field (key + fields).
_TSN_COL = {
    "PR": "PP", "PM": "POST_MILE", "HG": "HG", "City Code": "CITY_CODE", "R/U": "RU",
    "INT Type": "TY_INT", "Control Type": "TY_CT", "Lighting": "LT_TY",
    "ML Mastarm": "MAIN_SM", "ML Left Chan": "MAIN_LC", "ML Right Chan": "MAIN_RC",
    "ML Traffic Flow": "MAIN_TF", "ML Num Lanes": "MAIN_NL", "Description": "DESCRIPTION",
    "CS Mastarm": "CS_SM", "CS Left Chan": "CS_LC", "CS Right Chan": "CS_RC",
    "CS Traffic Flow": "CS_TF", "CS Num Lanes": "CS_NL", "Date of Record": "DATE_REC",
    # added columns — mapped BY REPORT POSITION (each report column compared to the same
    # column in the other report; user decision 2026-06-24). The mainline/cross eff-date
    # (next to the attrs) maps to TSN's historical/geometry EFF_DATE_ML / CROSS_BEGIN_DATE
    # and the Int St eff-date to TSN's EFF_DATE (both prints show those there). Since the
    # July 2026 update Int St is the structural one — TSN stores a bulk refresh stamp
    # ('2022-01-01') where TSMIS keeps the historical date.
    "INT Type Eff-Date": "EFF_DATE_INT", "Control Type Eff-Date": "EFF_DATE_CT",
    "Lighting Eff-Date": "EFF_DATE_LT", "ML Eff-Date": "EFF_DATE_ML",
    "CS Eff-Date": "CROSS_BEGIN_DATE", "Main Line Length": "MAIN_OVERRIDE",
    "Int St Eff-Date": "EFF_DATE",
    "Intrte Route": "CROSS_ROUTE_NAME", "Intrte PM Prefix": "CROSS_PM_PREFIX",
    "Intrte Postmile": "CROSS_POSTMILE", "Intrte PM Suffix": "CROSS_PM_SUFFIX",
    "Xing Line Lgth": "X_CROSS_OVERRIDE",
}
TSN_RAW_HEADER = (
    "PP", "POST_MILE", "LOCATION", "DATE_REC", "HG", "CITY_CODE", "RU",
    "EFF_DATE_INT", "TY_INT", "EFF_DATE_CT", "TY_CT", "EFF_DATE_LT", "LT_TY",
    "EFF_DATE_ML", "MAIN_SM", "MAIN_LC", "MAIN_RC", "MAIN_TF", "MAIN_NL",
    "X_CROSS_OVERRIDE", "MAIN_EFF_DATE", "MAIN_ADT", "DESCRIPTION",
    "MAIN_OVERRIDE", "CROSS_BEGIN_DATE", "CS_SM", "CS_LC", "CS_RC", "CS_TF",
    "CS_NL", "EFF_DATE", "CROSS_ADT", "CROSS_ROUTE_NAME", "CROSS_PM_PREFIX",
    "CROSS_POSTMILE", "CROSS_PM_SUFFIX",
)
# Consolidated-TSMIS VALUE position for each shared field (Route at 0; some header
# labels are still shifted against their values — the eff-date sits under each block's
# type label — so position, not label, stays authoritative; verified statewide against
# the July 2026 export, 16,200 paired rows).
_TSMIS_POS = {
    "PR": 1, "PM": 2, "HG": 6, "City Code": 7, "R/U": 8, "INT Type": 10,
    "Control Type": 12, "Lighting": 14, "ML Mastarm": 16, "ML Left Chan": 17,
    "ML Right Chan": 18, "ML Traffic Flow": 19, "ML Num Lanes": 20, "Description": 21,
    "CS Mastarm": 24, "CS Left Chan": 25, "CS Right Chan": 26, "CS Traffic Flow": 27,
    "CS Num Lanes": 28, "Date of Record": 5,
    "INT Type Eff-Date": 9, "Control Type Eff-Date": 11, "Lighting Eff-Date": 13,
    "ML Eff-Date": 15, "Main Line Length": 22, "CS Eff-Date": 23,
    "Int St Eff-Date": 29, "Intrte Route": 31, "Intrte PM Prefix": 32,
    "Intrte Postmile": 33, "Intrte PM Suffix": 34, "Xing Line Lgth": 35,
}
_TSMIS_ROUTE_POS = 4                       # consolidated "Location" column ("12 ORA 001")
# CMP-AUD-034: the EXACT consolidated header (['Route'] + the 35 source columns),
# bound exactly — _tsmis_row reads every field BY POSITION, so the old "len==36 and
# last=='Xing Line Lgth'" gate let a junk-relabelled/block-shifted header mis-map
# every field. TWO valid site editions are accepted (the value POSITIONS are
# identical across both — only the LABELS differ — so the by-position reader is
# unaffected; a pre-July-2026 37-column layout or any real column move still matches
# neither and is refused):
#   * CURRENT (the 2026-07-17 build): the site corrected its long-misaligned header
#     labels — 'P'->'PP', 'S'->'PS', the INT Type/INT Eff-Date labels swapped to sit
#     over their own values, 'Ctrl T'->'Ctrl T Eff-Date', 'Xing P/S'->'Int PS'.
#   * LEGACY (7.8/7.9): the prior labels, kept for backward compatibility.
_TSMIS_HEADER = [
    "Route", "PP", "Post Mile", "PS", "Location", "Date of Record", "H/G",
    "City Code", "R/U", "INT Eff-Date", "INT Type", "Ctrl T Eff-Date", "Ctrl Type",
    "Light Eff-Date", "Light T/Y", "ML Eff-Date", "ML S/M", "ML L/C", "ML R/C",
    "ML T/P", "ML N/L", "Description", "Main Line Lgth", "Inter Eff-Date",
    "Inter S", "Inter L", "Inter R", "Inter T", "Inter N", "Int St Eff-Date",
    "Intrte S", "Intrte Route", "Intrte Post", "Intrte Mile", "Int PS",
    "Xing Line Lgth"]
_TSMIS_HEADER_LEGACY = [
    "Route", "P", "Post Mile", "S", "Location", "Date of Record", "H/G",
    "City Code", "R/U", "INT Type", "INT Eff-Date", "Ctrl T", "Ctrl Type",
    "Light Eff-Date", "Light T/Y", "ML Eff-Date", "ML S/M", "ML L/C", "ML R/C",
    "ML T/P", "ML N/L", "Description", "Main Line Lgth", "Inter Eff-Date",
    "Inter S", "Inter L", "Inter R", "Inter T", "Inter N", "Int St Eff-Date",
    "Intrte S", "Intrte Route", "Intrte Post", "Intrte Mile", "Xing P/S",
    "Xing Line Lgth"]
_HEADER_LEN = len(_TSMIS_HEADER)           # 36 (kept for any positional references)


_header_ok = ctc.exact_consolidated_header_ok(_TSMIS_HEADER, _TSMIS_HEADER_LEGACY)


# --------------------------------------------------------------------------- #
# normalization
# --------------------------------------------------------------------------- #
def _split_route(tok):
    """Split a LOCATION token into (base_route, route_suffix):
    '12 ORA 210U' -> ('210', 'U'); '12 ORA. 210' -> ('210', '').

    A California route name can carry an alpha route SUFFIX (e.g. S/U — the printed
    report's "S" column) that TSN keeps but TSMIS often omits. Keying on the BASE
    route lets the same intersection still pair across that label difference, while
    the suffix is surfaced as the compared 'Route Suffix' column — so a suffix-only
    difference is flagged there (TSN 'U' vs TSMIS blank) rather than the rows being
    dropped to one-sided OR silently merged."""
    t = ("" if tok is None else str(tok)).strip().upper().replace("-", " ")
    parts = t.split()
    last = parts[-1] if parts else ""
    m = re.fullmatch(r"(\d+)([A-Z]?)", last)
    return (f"{int(m.group(1)):03d}", m.group(2)) if m else (last, "")


def _norm_route(tok):
    """The base route number (route suffix stripped) — the row key. See _split_route."""
    return _split_route(tok)[0]


def _dist_cnty(loc):
    """LOCATION '12 ORA 210U' / '04 CC. 004' -> ('12', 'ORA'/'CC')."""
    parts = ("" if loc is None else str(loc)).strip().upper().replace("-", " ").split()
    dist = parts[0] if parts else ""
    cnty = parts[1].rstrip(".") if len(parts) >= 2 else ""
    return (f"{int(dist):02d}" if dist.isdigit() else dist), cnty


def _decimal_pm(pm_raw):
    """The ID-79 oracle's Decimal-canonical numeric postmile as text —
    '005.870' -> '5.87', '001.000' -> '1', '0.000' -> '0'. Homed in
    compare_tsn_common since CMP-AUD-006 (Ramp Detail's identity now shares
    it); this name stays for the loaders and the golden canary."""
    return ctc.decimal_pm(pm_raw)


def _raw_text(v):
    return "" if v is None else str(v)


def _physical_id_key(base_route, county, pp_raw, pm_raw, claims, source_hint):
    """The accepted ID-79 PhysicalKey (CMP-AUD-045): canonical identity is
    exactly `(base Route, County, complete PP, numeric Post Mile)` — the
    complete PP is PART of identity (six real within-county groups carry
    distinct PPs at one numeric PM), rendered as the canonical postmile
    component `PP + Decimal-canonical PM` (e.g. "R5.87"); the route SUFFIX
    stays a compared column and conserved claim, never a key component. The
    key's str payload is the normalized PM the sheets display; a row without a
    usable county or postmile refuses loudly."""
    pm_display = _norm_pm(pm_raw)
    pp = _raw_text(pp_raw).strip().upper()
    numeric = _decimal_pm(pm_raw)
    if not county:
        raise ValueError(f"Intersection Detail row (route {base_route}, PM "
                         f"{pm_display}) has no usable county in {source_hint} "
                         "— cannot key it to a physical location")
    if not numeric:
        raise ValueError(f"Intersection Detail row (route {base_route}, "
                         f"{source_hint}) has no usable numeric postmile")
    identity = cc.make_physical_identity(
        base_route, county, f"{pp}{numeric}",
        tuple(cc.RawIdentityClaim(name, value) for name, value in claims),
        f"{base_route} / {county} / {pp}{numeric}")
    return cc.physical_key(pm_display, identity)


# PM + date canon shared with Ramp Detail, homed in compare_tsn_common (P5b/S04);
# iso_date also handles this report's 2-digit TSN year. Names kept so the loaders, the
# Report View, and the golden canary still resolve idt._norm_pm / idt._iso_date.
_norm_pm = ctc.norm_pm
_iso_date = ctc.iso_date


_BOOL = {"Y": "Y", "N": "N", "1": "Y", "0": "N"}


def _norm_bool(v):
    # None-safe like _norm_num: a numeric 0 must read as "0" (-> 'N'), not "" — the
    # `v or ""` coercion would drop a real 0 to blank and flag a phantom diff.
    s = ("" if v is None else str(v)).strip()
    return _BOOL.get(s.upper(), s)


def _norm_num(v):
    """Canonicalize a zero-padded number: '058'->'58', '9.560'->'9.56', '0.000'->'0'.
    Non-numeric values are returned unchanged (so e.g. a route name with letters is safe).
    A real numeric 0/0.0 canonicalizes to '0', NOT blank: `str(v or "")` coerces a
    numeric 0 to "" (0 is falsy), so TSN's numeric-0 intersecting-route postmile read as
    blank while TSMIS's text '0.000' read as '0' — a phantom 0-vs-blank diff. None/"" stay
    blank, so a genuine value-vs-missing difference still flags."""
    s = ("" if v is None else str(v)).strip()
    if not s or not re.fullmatch(r"-?\d+(\.\d+)?", s):
        return s
    neg, s = s.startswith("-"), s.lstrip("-")
    if "." in s:
        ip, fp = s.split("."); ip = ip.lstrip("0") or "0"; fp = fp.rstrip("0")
        s = ip + ("." + fp if fp else "")
    else:
        s = s.lstrip("0") or "0"
    return ("-" + s) if neg else s


def _norm_control_type(v):
    """Apply the TSN→TSNR control-type crosswalk: the legacy signal sub-types J–P
    (TSN) and TSMIS's combined "S" all fold into the single code "S" (the Signalized
    category TSMIS stores), so the sub-type split no longer reads as a difference. The
    compared cell therefore shows "S" wherever the crosswalk applied. Every other code
    is left as-is (both systems share A/B/C/D/E/F/G/H/I/R/Z)."""
    s = ("" if v is None else str(v)).strip().upper()
    return _SIGNALIZED_LABEL if s in _SIGNALIZED_CODES else _v(v)


def _v(x):
    return normalize_value(x)


def _project(field, raw):
    """Normalize one raw cell for `field` into the shared, comparable form."""
    if field in BOOLEAN_FIELDS:
        return _norm_bool(raw)
    if field == "Control Type":
        return _norm_control_type(raw)
    if field == "PM":
        return _norm_pm(raw)
    if field in DATE_FIELDS:
        return _iso_date(raw)
    if field in NUMERIC_FIELDS:
        return _norm_num(raw)
    return _v(raw)


# --------------------------------------------------------------------------- #
# loaders -> consolidated-shape rows ([route, *SHARED_HEADER])
# --------------------------------------------------------------------------- #
def _tsn_row(r, h):
    def g(name):
        i = h.get(name)
        return r[i] if i is not None and i < len(r) else None
    loc = g("LOCATION")
    base, route_suffix = _split_route(loc)
    district, county = _dist_cnty(loc)
    key = _physical_id_key(base, county, g("PP"), g("POST_MILE"), (
        ("route", base), ("route_suffix", route_suffix),
        ("location", _raw_text(loc)),
        ("postmile_prefix", _raw_text(g("PP"))),
        ("postmile", _raw_text(g("POST_MILE")))), f"LOCATION {loc!r}")
    derived = {"Route Suffix": route_suffix, KEY: key,
               "District": district, "County": county}
    return [base] + [derived[f] if f in derived
                     else _project(f, g(_TSN_COL[f]))
                     for f in SHARED_HEADER]


def require_tsn_raw_header(header):
    ctc.require_exact_raw_header(header, TSN_RAW_HEADER, REPORT_NAME)


def tsn_rows_from_raw(path):
    with ctc.exact_raw_rows(
            path, TSN_SHEET, TSN_RAW_HEADER, REPORT_NAME,
            required_nonblank=("LOCATION", "POST_MILE")) as (header, rows_in):
        h = {n: i for i, n in enumerate(header)}
        return [_tsn_row(r, h) for r in rows_in]


def _normalized_row(r):
    """Re-project one row from the normalized TSN-library sheet (v3: ['Route'] +
    SHARED_HEADER + the District/County sidecars) onto the shared shape,
    RE-APPLYING the field normalizations (control-type crosswalk, booleans,
    PM, date). The projections are idempotent on already-normalized values, so
    this is a no-op for a freshly-built library BUT repairs a STALE one.
    The ID-79 PhysicalKey is rebuilt from the row's base route + County sidecar
    + PR/PM columns (CMP-AUD-045 — the sidecars are no longer sliced away; the
    v3 shape already carries everything, so no library rebuild is needed)."""
    width = len(SHARED_HEADER) + 1
    vals = list(r)[:width + 2]                    # + District/County sidecars
    vals += [None] * (width + 2 - len(vals))
    base = _raw_text(vals[0]).strip()
    district, county = (_raw_text(vals[width]).strip(),
                        _raw_text(vals[width + 1]).strip().rstrip("."))
    pp = vals[1 + SHARED_HEADER.index("PR")]
    pm = vals[1 + SHARED_HEADER.index(KEY)]
    suffix_i = 1 + SHARED_HEADER.index("Route Suffix")
    key = _physical_id_key(base, county, pp, pm, (
        ("route", base), ("route_suffix", _raw_text(vals[suffix_i])),
        ("district", district), ("county", county),
        ("postmile_prefix", _raw_text(pp)), ("postmile", _raw_text(pm))),
        "the library's TSN County sidecar")
    out = [base] + [_project(f, vals[i + 1]) for i, f in enumerate(SHARED_HEADER)]
    out[1 + SHARED_HEADER.index(KEY)] = key
    out[1 + SHARED_HEADER.index("District")] = district
    out[1 + SHARED_HEADER.index("County")] = county
    return out


def _load_tsn(path):
    path = Path(path)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open {path.name}: {type(e).__name__}: {e}")
    try:
        if NORMALIZED_SHEET in wb.sheetnames:
            it = wb[NORMALIZED_SHEET].iter_rows(values_only=True)
            header = [_raw_text(c).strip() for c in (next(it, None) or ())]
            # CMP-AUD-033: bind the header to the exact ["Route"] + SHARED_HEADER
            # prefix + documented sidecars before reading BY POSITION. This
            # subsumes the CMP-AUD-045 pre-county-aware shape check (a pre-v3
            # library lacks the sidecars) and additionally refuses a reordered
            # or renamed shared header that would mis-map every column.
            ctc.require_shared_header_prefix(
                header, ["Route"] + SHARED_HEADER, _NORMALIZED_SIDECARS,
                path.name, REPORT_NAME)
            # CMP-AUD-037: a correctly-shaped library missing the in-workbook
            # marker predates the direct-path freshness gate — refuse it. The
            # library path already auto-rebuilds via D2.
            ctc.require_current_normalization(
                wb, path.name, NORMALIZATION_VERSION,
                "pre-v5: no in-workbook normalization marker")
            rows = [_normalized_row(r)
                    for r in it if r and any(c not in (None, "") for c in r)]
            return rows, True
    finally:
        wb.close()
    return tsn_rows_from_raw(path), True


def _tsmis_row_with(r, project):
    """One consolidated TSMIS row with `project(field, raw)` supplying the
    value projection. The 045 physical pairing key and the Location-derived
    provenance are IDENTICAL for every caller; only the value projection
    varies — `_project` for the vs-TSN comparison (cross-system crosswalks),
    a verbatim projection for the same-source PDF-vs-Excel flavor
    (CMP-AUD-067: crosswalks must not erase render differences between two
    TSMIS renders)."""
    def at(i):
        return r[i] if i < len(r) else None
    loc = at(_TSMIS_ROUTE_POS)
    base, route_suffix = _split_route(loc)
    district, county = _dist_cnty(loc)
    key = _physical_id_key(base, county, at(_TSMIS_POS["PR"]),
                           at(_TSMIS_POS[KEY]), (
        ("route", base), ("route_suffix", route_suffix),
        ("location", _raw_text(loc)),
        ("postmile_prefix", _raw_text(at(_TSMIS_POS["PR"]))),
        ("postmile", _raw_text(at(_TSMIS_POS[KEY]))),
        # The export's S column (position 3), conserved as a source claim —
        # NOT identity (the accepted ID-79 tuple carries no suffix).
        ("postmile_suffix", _raw_text(at(3)))), f"Location {loc!r}")
    derived = {"Route Suffix": route_suffix, KEY: key,
               "District": district, "County": county}
    return [base] + [derived[f] if f in derived
                     else project(f, at(_TSMIS_POS[f]))
                     for f in SHARED_HEADER]


def _tsmis_row(r):
    return _tsmis_row_with(r, _project)


def _load_tsmis(path):
    return load_consolidated_rows(
        path, TSMIS_SHEET,
        missing_sheet_hint="pick the consolidated TSMIS Intersection Detail workbook.",
        bad_header_msg="isn't a CONSOLIDATED Intersection Detail workbook in the "
                       "current (July 2026) site format — a leading 'Route' column "
                       "and the 'Xing Line Lgth' tail column are expected. "
                       "Consolidate a fresh post-update export; pre-update exports "
                       "used the old 36-column layout, which this version doesn't "
                       "compare.",
        header_ok=_header_ok,
        row_transform=_tsmis_row)


# --------------------------------------------------------------------------- #
# Notes sheet — documents every normalization applied (so a match is read as
# "equal after the stated normalization", not raw equality) and comments on the
# columns that differ wholesale. Nothing is suppressed: all shared fields are
# compared and counted.
# --------------------------------------------------------------------------- #
def _write_notes_sheet(wb):
    ws = wb.create_sheet("Notes")
    ws.sheet_properties.tabColor = "ED7D31"
    write_only = getattr(wb, "write_only", False)
    title = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    head = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color="1F3864")
    sec_fill = PatternFill("solid", start_color="0070C0")
    body = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)

    def cell(value, font=body, f=None, align=None):
        if not write_only:
            return value
        c = WriteOnlyCell(ws, value=value)
        c.font = font
        if f:
            c.fill = f
        if align:
            c.alignment = align
        return c

    def section(text):
        ws.append([cell(text, head, f=sec_fill)])

    def note(text):
        ws.append([cell(text, body, align=wrap)])

    ws.column_dimensions["A"].width = 110
    ws.append([cell("Intersection Detail — TSMIS vs TSN: comparison notes", title, fill)])
    note("This is a MECHANICAL field-by-field diff. EVERY column present in both systems "
         "is compared and counted — nothing is hidden. A few columns differ structurally "
         "(noted below); they are still flagged, and the reasons are noted here "
         "(commentary belongs to the reader, not to suppressing a column). The TSMIS side "
         "must be a JULY-2026-or-later export: the site update of 2026-07 reshaped the "
         "report (35 columns), and pre-update workbooks are refused rather than mis-read.")

    section("NORMALIZATIONS APPLIED  (these can make raw-different values compare EQUAL — "
            "a match means 'equal after the normalization named here', not that the source "
            "cells were byte-identical)")
    note("• Control Type — CROSSWALK (per the TSNR/MIRE reference 'TSNR - Intersection "
         "Control and Geometry Type'): TSN records signalized intersections under the legacy "
         "signal sub-types J/K/L/M/N/P (pretimed / semi- / full-actuated, 2- vs multi-phase); "
         "TSNR/TSMIS collapses them into ONE category stored as the code 'S'. Both sides' "
         "signalized codes are normalized to that single code 'S' (the Signalized category). "
         "HOW TO SEE IT: wherever the Control Type cell reads 'S' (a single code that may be "
         "a folded TSN J–P), the crosswalk was applied — a TSN J–P matched a TSMIS S. Every other control code "
         "(A/B/C/D/E/F/G/H/I/R/Z) is shared and compared unchanged. INT Type needs no "
         "crosswalk — both systems share the F/M/S/T/Y/Z/R codes.")
    note("• Boolean encoding — Lighting, ML Mastarm, ML Right Chan, CS Mastarm, CS Right "
         "Chan are Y/N on BOTH sides since the July 2026 update. Pre-update TSMIS data "
         "stored 1/0; the Y≡1 / N≡0 normalization still folds any legacy-encoded value, so "
         "only a genuine change flags (not the encoding).")
    note("• Postmile (PM) — leading zeros and spaces are stripped so the same postmile pairs "
         "across formatting (TSN ' 004.901' ≡ TSMIS '004.901' ≡ an older '4.901'). PM is "
         "the row key.")
    note("• Route suffix — a California route name can carry an alpha suffix (e.g. S/U). "
         "Rows are matched on the BASE route number so the same intersection pairs either "
         "way; the suffix is compared in the 'Route Suffix' column. Since July 2026 the "
         "TSMIS Location carries the suffix too ('11 IMP 008U'), so this column now flags "
         "only genuine suffix disagreements, not the old TSMIS-blank gap.")
    note("• Numeric zero-padding — Main Line Length ('58' ≡ '058'), the intersecting-route "
         "number and postmile ('9.560' ≡ '9.56'), and Xing Line Lgth (TSN's "
         "X_CROSS_OVERRIDE) canonicalize before comparing.")
    note("• Quote characters are NOT normalized — they compare literally. Both systems "
         "write quoted street letters as a doubled apostrophe (''F'' ST) on almost every "
         "such row; where one side instead stores a real quotation mark (\"F\" ST) the "
         "Description flags. The two forms print near-identically, so the evidence images "
         "label such a difference explicitly — but the sides genuinely store different "
         "characters (statewide census 2026-07: exactly one such row, KER 046 @ 50.904).")

    section("COLUMNS THAT STILL DIFFER STRUCTURALLY  (compared and counted like any other — "
            "the difference is systematic, explained here, NOT a per-intersection data error)")
    note("• Int St Eff-Date — TSN stores a BULK refresh stamp ('2022-01-01' on ~99% of rows) "
         "where TSMIS keeps the historical date, so this column differs on nearly every "
         "matched row — the one wholesale-structural column left. Read the count as the "
         "convention, not as thousands of corrections.")
    note("• ML Eff-Date (~12% of rows) and CS Eff-Date (~3%) — TSN tracks the most recent "
         "mainline/cross-street resurvey date where TSMIS keeps the original geometry date "
         "(e.g. TSMIS '1964-01-01' vs TSN '1998-08-28'). Systematic direction, moderate "
         "volume; red on the Report View but kept out of its Major count.")
    note("• ML / CS attributes — a small completeness gap remains (~1% of rows): TSMIS "
         "leaves an attribute blank where TSN carries a value. Compared and counted; the "
         "old ~37% cross-street gap was closed by the July 2026 update.")
    note("• WHAT THE JULY 2026 UPDATE FIXED (expect these counts to be near zero now): "
         "Date of Record and the INT Type / Control / Lighting eff-dates now match TSN on "
         "≥99.9% of rows — the old wholesale refresh-vs-record difference AND the old "
         "systematic ~1-day offset are both gone; the remaining flags in those columns are "
         "GENUINE conflicts. 'Xing Line Lgth' (TSN X_CROSS_OVERRIDE) is newly exported by "
         "TSMIS and newly compared.")
    note("• Intersecting-route block (Intrte Route / PM Prefix / Postmile / PM Suffix) + Main "
         "Line Length — also compared. The intersecting route is mostly blank on both (only a "
         "few hundred intersections cross another state route); differences are genuine "
         "where present.")
    note("• Nothing is greyed-out or shown-but-not-counted — under position alignment every shared "
         "column is compared and counted. (TSMIS's blank route-suffix 'S' / 'Intrte S' stubs are "
         "omitted; TSN's MAIN_EFF_DATE — the second ML eff-date TSMIS no longer exports — and its "
         "ADT columns MAIN_ADT / CROSS_ADT have no TSMIS counterpart and aren't compared — they "
         "appear, for reference, only on the Report View.)")

    section("REPORT VIEW  (a second sheet — the printed two-line record, for visual inspection)")
    note("• The 'Report View' tab replicates the printed Intersection Detail record (two physical "
         "lines per intersection) and renders EVERY difference in red — the structural date "
         "columns included — so the page can be eyeballed straight against the source PDF. Per "
         "record it shows two counts: 'Major' = genuine conflicts (the three structural date "
         "columns — Int St / ML / CS Eff-Date — and the route suffix are excluded so they don't "
         "drown out the real conflicts; Date of Record and the INT/Control/Lighting eff-dates "
         "COUNT now that they match structurally); 'Diffs' = every difference. TSN-only "
         "reference columns (ML 2nd Eff-Date, the ADT pair) appear there in blue.")
    note("Rows are keyed on Route + Postmile (PM).")
    return ws


_SCHEMA = CompareSchema(
    report_name=REPORT_NAME,
    header=SHARED_HEADER,
    side_a="TSMIS",
    side_b="TSN",
    id_noun="intersection",
    id_noun_plural="intersections",
    pair_noun="postmile",
    sides_noun="systems",
    date_fields=DATE_FIELDS,
    data_widths={"Description": 26, "Date of Record": 11},
    cmp_widths={"Description": 30, "Date of Record": 12},
    one_sided_note_extra=" (intersections one system lists at a postmile the other doesn't)",
    key_field=KEY_FIELD,
    context_fields=CONTEXT_FIELDS,  # () — position-aligned, nothing suppressed or greyed
    legend_writer=_write_notes_sheet,
)


# --------------------------------------------------------------------------- #
# Report View — a two-line replica of the printed report, comparison-coloured
# --------------------------------------------------------------------------- #
# The printed Intersection Detail record is TWO physical lines; this sheet mirrors
# that (row 1 = MAINLINE side, row 2 = CROSS-STREET side) so the comparison reads
# like the report. Mainline and intersecting blocks are PARALLEL (same attributes)
# so they share the middle columns; line-1-only fields sit on the mainline row,
# line-2-only on the cross row. Colour: RED = any difference — a genuine ("Major")
# discrepancy AND a structural date difference both read red (user request
# 2026-06-24: all date discrepancies are red); BLUE = TSN-only column; AMBER =
# TSMIS-only column. Two per-record counts head each record: Major = genuine
# conflicts (the structural columns below are excluded so they don't drown out the
# real differences), Diffs = every difference. Identity is repeated on both
# physical rows so a filter keeps the 2-row records together — the streaming
# workbook can't vertically merge cells.
_RV_ONE = {"LOC": "LOCATION", "ML2": "MAIN_EFF_DATE",
           "ADT": "MAIN_ADT", "CADT": "CROSS_ADT"}
_RV_DATEONE = ("ML2",)   # MAIN_EFF_DATE (the 2nd ML eff-date TSMIS no longer exports)
# The structural columns (user decision 2026-07-08, the data-driven soft set): Int St
# Eff-Date (TSN's bulk '2022-01-01' stamp vs TSMIS's historical date, ~99% differ),
# ML / CS Eff-Date (TSN tracks a later resurvey date, ~12% / ~3%). Date of Record and
# the INT/Control/Lighting eff-dates now match TSN structurally (the July 2026 export
# update; the old ~1-day offset is gone), so their differences are GENUINE and count
# as Major like any attribute conflict.
_RV_SOFT_ALWAYS = ("ML Eff-Date", "CS Eff-Date", "Int St Eff-Date")
_RV_AUX = ("Major", "Diffs", "Route")       # frozen-left aux columns
# Report grid — column SHARING matches the printed report: DESCRIPTION spans under
# LOCATION, LINE LGTH under R/U, the INTERSECTING block under INT/CONTROL/LIGHTING,
# INT ST + INTERSECTING ROUTE + XING under MAINLINE. Each entry is
# (g1, l1, s1, g2, l2, s2): line-1 group/label/spec stacked over line-2's. spec=(kind,ref).
_RV_GRID = [
    # Route Suffix sits with the route (frozen, immediately right of the aux Route column).
    # Compared like every other cmp cell — a TSN-U-vs-TSMIS-blank suffix difference flags
    # RED — but classified 'soft' (see _rv_classify) so the systematic gap doesn't inflate
    # the per-record Major count.
    ("", "SFX", ("cmp", "Route Suffix"), "", "", ("blank", None)),
    ("", "P", ("cmp", "PR"), "", "", ("blank", None)),
    ("", "POST MILE", ("pm", None), "", "", ("blank", None)),
    ("", "S", ("blank", None), "", "", ("blank", None)),
    ("", "LOCATION", ("loc", None), "", "DESCRIPTION", ("cmp", "Description")),
    ("", "DATE OF REC", ("cmp", "Date of Record"), "", "", ("blank", None)),
    ("", "H/G", ("cmp", "HG"), "", "", ("blank", None)),
    ("", "CITY", ("cmp", "City Code"), "", "", ("blank", None)),
    ("", "R/U", ("cmp", "R/U"), "*MAIN*", "LINE LGTH", ("cmp", "Main Line Length")),
    ("* INT *", "EFF-DATE", ("cmp", "INT Type Eff-Date"), "* INTERSECTING *", "EFF-DATE", ("cmp", "CS Eff-Date")),
    ("* INT *", "T/Y", ("cmp", "INT Type"), "* INTERSECTING *", "S/M", ("cmp", "CS Mastarm")),
    ("* CONTROL *", "EFF-DATE", ("cmp", "Control Type Eff-Date"), "* INTERSECTING *", "L/C", ("cmp", "CS Left Chan")),
    ("* CONTROL *", "T/Y", ("cmp", "Control Type"), "* INTERSECTING *", "R/C", ("cmp", "CS Right Chan")),
    ("* LIGHTING *", "EFF-DATE", ("cmp", "Lighting Eff-Date"), "* INTERSECTING *", "T/F", ("cmp", "CS Traffic Flow")),
    ("* LIGHTING *", "T/Y", ("cmp", "Lighting"), "* INTERSECTING *", "N/L", ("cmp", "CS Num Lanes")),
    ("* MAINLINE *", "EFF-DATE", ("cmp", "ML Eff-Date"), "* INT ST *", "EFF-DATE", ("cmp", "Int St Eff-Date")),
    ("* MAINLINE *", "S/M", ("cmp", "ML Mastarm"), "*INT ROUTE*", "RTE NO", ("cmp", "Intrte Route")),
    ("* MAINLINE *", "L/C", ("cmp", "ML Left Chan"), "*INT ROUTE*", "P", ("cmp", "Intrte PM Prefix")),
    ("* MAINLINE *", "R/C", ("cmp", "ML Right Chan"), "*INT ROUTE*", "POST MI", ("cmp", "Intrte Postmile")),
    ("* MAINLINE *", "T/F", ("cmp", "ML Traffic Flow"), "*XING*", "P/S", ("cmp", "Intrte PM Suffix")),
    ("* MAINLINE *", "N/L", ("cmp", "ML Num Lanes"), "*XING*", "LINE LGTH", ("cmp", "Xing Line Lgth")),
    ("TSN only", "ML 2nd EFF", ("tn", "ML2"), "TSN only", "", ("blank", None)),
    ("TSN only", "ML ADT", ("tn", "ADT"), "TSN only", "CS ADT", ("tn", "CADT")),
]
# (normal, ALT) fill hex pairs — whole-record alternation across every cell type. A record's
# neutral cells are uniformly WHITE (normal) or GREY (alt) so each record reads as one solid
# zebra band: white is applied to EVERY neutral cell of the record (blanks included), never
# mixed with grey within a record (the patchy-white-gap bug). 'soft' (date differences)
# shares the hard RED palette — every date discrepancy renders red (user request 2026-06-24)
# — while staying out of the Major count (see _rv_classify).
_RV_FILLS = {"hard": ("F8D4D4", "E8B6B6"), "soft": ("F8D4D4", "E8B6B6"),
             "tn": ("DCE5F3", "B7CCE7"), "tm": ("F8E4CF", "E3C9A2"),
             "id": ("FFFFFF", "CFD6DE"), "count": ("FFFFFF", "CFD6DE"), "eq": ("FFFFFF", "CFD6DE")}
_RV_FONTCOL = {"hard": "9C0006", "soft": "9C0006", "tn": "163A63", "tm": "7A431A"}
# Hover-comments on the normalized headers (so a match reads as "equal after this rule").
_RV_COMMENTS = {
    "Control Type": "NORMALIZED: TSN's signal sub-types J-P and TSMIS's 'S' all fold to 'S' "
                    "(signalized) per the TSNR crosswalk, so the sub-type split doesn't flag.",
    "Lighting": "NORMALIZED: both sides store Y/N; a legacy TSMIS 1/0 folds to Y/N.",
    "ML Mastarm": "NORMALIZED: Y/N both sides; legacy 1/0 folds to Y/N.",
    "ML Right Chan": "NORMALIZED: Y/N both sides; legacy 1/0 folds to Y/N.",
    "CS Mastarm": "NORMALIZED: Y/N both sides; legacy 1/0 folds to Y/N.",
    "CS Right Chan": "NORMALIZED: Y/N both sides; legacy 1/0 folds to Y/N.",
    "Main Line Length": "NORMALIZED: zero-padding ignored (TSMIS '58' = TSN '058').",
    "Intrte Postmile": "NORMALIZED: trailing zeros ignored (TSMIS '9.560' = TSN '9.56').",
    "Xing Line Lgth": "NORMALIZED: zero-padding ignored. TSN's X_CROSS_OVERRIDE - newly "
                      "exported by TSMIS (July 2026) and compared.",
    "Date of Record": "Both sides store the historical record date since the July 2026 "
                      "export update - a difference here is a GENUINE conflict and "
                      "counts as Major.",
    "ML Eff-Date": "TSN tracks the most recent mainline resurvey date; TSMIS keeps the "
                   "original on ~12% of rows - structural. Shown RED, not counted as Major.",
    "CS Eff-Date": "TSN tracks the most recent cross-street resurvey date; TSMIS keeps the "
                   "original on ~3% of rows - structural. Shown RED, not counted as Major.",
    "Int St Eff-Date": "TSN stores a bulk refresh stamp ('2022-01-01' on ~99% of rows); "
                       "TSMIS the historical date - structural. Shown RED, not counted "
                       "as Major.",
}


def _rv_classify(field, tm, tn):
    """Classify a DIFFERING cell. 'soft' = a structural difference (Int St / ML / CS
    Eff-Date — see _RV_SOFT_ALWAYS — or the route suffix): it renders RED like a
    genuine conflict but is kept OUT of the per-record Major count. 'hard' = a genuine
    'Major' discrepancy — since the July 2026 export update that includes Date of
    Record and the INT/Control/Lighting eff-dates (they match TSN structurally now;
    the old ~1-day-offset tolerance is retired). Both soft and hard count toward
    Diffs. (`tm`/`tn` stay in the signature for parity with the grid's cmp cells.)"""
    del tm, tn
    if field == "Route Suffix":
        # Matched on the base route; a suffix disagreement is a labeling gap, not a
        # geometry conflict: red, but not Major.
        return "soft"
    if field in _RV_SOFT_ALWAYS:
        return "soft"
    return "hard"


def _tsn_onesided(path):
    """Raw TSN one-sided columns (the second ML eff-date TSMIS no longer exports +
    the ADT counts) + Location, aligned to the rows `tsn_rows_from_raw` yields.
    Returns None for a normalized-library workbook (those columns aren't stored
    there) — the replica then shows the TSN-only cells blank."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        # An unreadable input silently degraded the Report View to blanks --
        # the WHY goes to the log (compare_core output is untouched).
        log.info("report view: TSN one-sided read failed (%s: %s)",
                 type(e).__name__, str(e).splitlines()[0] if str(e) else "")
        return None
    try:
        if NORMALIZED_SHEET in wb.sheetnames:
            return None
        sn = TSN_SHEET if TSN_SHEET in wb.sheetnames else wb.sheetnames[0]
        it = wb[sn].iter_rows(values_only=True)
        hdr = next(it, None) or []
        h = {str(n).strip(): i for i, n in enumerate(hdr) if n is not None}
        out = []
        for r in it:
            if not row_has_data(r):
                continue
            out.append({k: ("" if h.get(col) is None or h[col] >= len(r) or r[h[col]] is None
                            else str(r[h[col]]).strip())
                        for k, col in _RV_ONE.items()})
        return out
    finally:
        wb.close()


def _tsmis_locations(path):
    """Consolidated TSMIS 'Location' (pos 4), aligned to the rows `_load_tsmis` yields."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        log.info("report view: TSMIS locations read failed (%s: %s)",
                 type(e).__name__, str(e).splitlines()[0] if str(e) else "")
        return []
    try:
        if TSMIS_SHEET not in wb.sheetnames:
            return []
        it = wb[TSMIS_SHEET].iter_rows(values_only=True)
        next(it, None)
        out = []
        for r in it:
            if not row_has_data(r):
                continue
            out.append("" if len(r) <= 4 or r[4] is None else str(r[4]).strip())
        return out
    finally:
        wb.close()


def _write_report_view(wb, ctx, tsn_one, tm_loc):
    """Append the two-line 'Report View' — a faithful replica of the printed
    Intersection Detail record (two physical lines per intersection) — to the
    streaming comparison workbook. Column SHARING mirrors the report: DESCRIPTION
    spans under LOCATION, LINE LGTH under R/U, the INTERSECTING block under
    INT/CONTROL/LIGHTING, INT ST + INTERSECTING ROUTE + XING under MAINLINE. The
    4-row header is the report's two stacked header blocks (line-1 group/label over
    line-2 group/label). Identity (Major/Diffs/Route + P/PostMile/S/Location) repeats
    on both physical rows so a filter keeps the 2-row records intact — the streaming
    workbook can't vertically merge the data cells (header cells it can, below)."""
    sc = ctx["sc"]
    events = ctx.get("events")
    rows_a, rows_b = ctx["rows_a"], ctx["rows_b"]
    # F1: reuse the pairing run_compare just computed (identical inputs for this
    # family: key_field, has_route=True, no key_normalizer); recompute only under
    # an older core that doesn't pass it.
    ka, kb, union = ctx.get("keys_a"), ctx.get("keys_b"), ctx.get("union")
    if ka is None or kb is None or union is None:
        is_cancelled = events.is_cancelled if events is not None else None
        ka = keys_for(
            rows_a, True, key_field=sc.key_field,
            is_cancelled=is_cancelled)
        kb = keys_for(
            rows_b, True, key_field=sc.key_field,
            is_cancelled=is_cancelled)
        pairing = pair_occurrences_by_similarity(
            sc, rows_a, rows_b, ka, kb, True, events)
        if pairing.pairing_quality != "exact":
            raise ValueError(
                "Report View cannot discard capped duplicate-pairing state")
        ka, kb = pairing.keys_a, pairing.keys_b
        union = union_keys(ka, kb, is_cancelled)
    # This rollup re-lays-out every record as two styled physical rows (~2x the union),
    # the slowest stretch of the largest comparison. Narrate it (header + per-N progress
    # below) so it isn't a silent multi-minute gap that reads as a freeze.
    if events is not None:
        events.on_log(f"  Building the Report View tab ({len(union):,} records)…")
    amap = {k: i for i, k in enumerate(ka)}
    bmap = {k: j for j, k in enumerate(kb)}
    field_index = {name: i for i, name in enumerate(sc.header)}
    fi = {name: 1 + i for i, name in enumerate(sc.header)}      # +1 for leading route col
    NA, NG = len(_RV_AUX), len(_RV_GRID)
    NC = NA + NG

    def aval(row, name):
        if row is None:
            return ""
        v = row[fi[name]]
        return "" if v is None else str(v).strip()

    Fn = lambda **k: Font(name="Consolas", **{"size": 8.5, **k})
    fill = lambda c: PatternFill("solid", fgColor=c)
    HEAD, GRP = fill("21344F"), fill("3A5688")
    thin = Side(style="thin", color="D2D2D2")
    med = Side(style="medium", color="51607A")        # strong between-record divider
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctrW = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")
    _FONTS = {"hard": dict(color=_RV_FONTCOL["hard"], bold=True),
              "soft": dict(color=_RV_FONTCOL["soft"], bold=True),
              "tn": dict(color=_RV_FONTCOL["tn"]), "tm": dict(color=_RV_FONTCOL["tm"]),
              "id": dict(bold=True), "count": dict(bold=True)}
    # F1: openpyxl styles are immutable and shareable — build every distinct
    # Border/Fill/Font ONCE instead of ~3 fresh objects per cell (~1.3M cells on
    # the statewide rollup). The serialized workbook is identical.
    _BD_NORM = Border(left=thin, right=thin)
    _BD_BOTTOM = Border(left=thin, right=thin, bottom=med)
    _FILL_CACHE = {(st, a): fill(cols[1 if a else 0])
                   for st, cols in _RV_FILLS.items() for a in (False, True)}
    _FONT_CACHE = {st: Fn(**kw) for st, kw in _FONTS.items()}
    _FONT_DEFAULT = Fn()

    ws = wb.create_sheet("Report View")
    ws.sheet_properties.tabColor = "21344F"
    ws.freeze_panes = "I5"      # MUST precede the streamed rows in write-only mode; keeps
                                # Major/Diffs/Route + Route Suffix + P/PostMile/S/Location
                                # + the 4 header rows (one wider than before: the SFX column)

    def value(spec, ra, rb, one):
        """Resolve a grid spec to (text, status). 'cmp' compares the normalized TSMIS
        and TSN cells (eq / soft / hard); 'tn' is a TSN-only column; everything else
        (blank / pm / loc handled in the row loop) is non-counting."""
        kind, ref = spec
        if kind == "tn":
            v = one.get(ref, "") if one else ""
            return (_iso_date(v) if ref in _RV_DATEONE else v, "tn")
        if kind == "tm":
            return (aval(ra, ref), "tm")
        if kind == "cmp":
            cell = compared_cell(sc, field_index[ref], ra, rb, off=1)
            if not cell.asserting or cell.equal:
                return (cell.display, "eq")
            tm, tn = cell.display_a, cell.display_b
            return (f"{tm or '·'} ≠ {tn or '·'}", _rv_classify(ref, tm, tn))
        return ("", "blank")

    def woc(val, status, alt, *, bottom=False, align=None):
        """A streamed data cell. Every status carries a (normal, ALT) band pair so a whole
        record alternates as one solid zebra band (white record vs grey record); an unknown
        status falls to 'eq' so a blank cell takes its record's band, not a stray shade."""
        c = set_safe_literal_cell(WriteOnlyCell(ws), val)
        c.alignment = align or (lft if status == "id" else ctr)
        c.border = _BD_BOTTOM if bottom else _BD_NORM
        c.fill = _FILL_CACHE.get((status, bool(alt)),
                                 _FILL_CACHE[("eq", bool(alt))])
        c.font = _FONT_CACHE.get(status, _FONT_DEFAULT)
        return c

    def hcell(val, fillc, font, align, comment_ref=None):
        """A streamed header cell, optionally carrying the normalization hover-comment."""
        c = WriteOnlyCell(ws, value=val)
        c.fill = fillc; c.font = font; c.alignment = align; c.border = bd
        t = _RV_COMMENTS.get(comment_ref) if comment_ref else None
        if t:
            cm = Comment(t, "TSMIS vs TSN"); cm.width = 250; cm.height = 130
            c.comment = cm
        return c

    # ---- 4-row header (the report's two stacked header blocks) ----
    g1 = [col[0] for col in _RV_GRID]
    g2 = [col[3] for col in _RV_GRID]

    def group_row(groups):
        """One header row of GROUP cells: the leftmost of each run carries the label and
        every cell in a non-empty run shares the GRP fill; empty groups stay plain."""
        cells, i = [], 0
        while i < NG:
            g, j = groups[i], i
            while j < NG and groups[j] == g:
                j += 1
            for k in range(i, j):
                if g:
                    cells.append(hcell(g if k == i else "", GRP,
                                       Fn(bold=True, color="FFFFFF", size=7.5), ctr))
                else:
                    cells.append(hcell("", PatternFill(), Fn(), ctr))
            i = j
        return cells

    aux_white = Fn(bold=True, color="FFFFFF", size=8)
    blank_dark = Fn()
    ws.append([hcell(lab, HEAD, aux_white, ctrW) for lab in _RV_AUX] + group_row(g1))
    ws.append([hcell("", HEAD, blank_dark, ctrW) for _ in _RV_AUX] +
              [hcell(col[1], HEAD, Fn(bold=True, color="FFFFFF", size=7.5), ctrW, col[2][1])
               for col in _RV_GRID])
    ws.append([hcell("", HEAD, blank_dark, ctrW) for _ in _RV_AUX] + group_row(g2))
    ws.append([hcell("", HEAD, blank_dark, ctrW) for _ in _RV_AUX] +
              [hcell(col[4], HEAD, Fn(color="BBD0EC", size=7.5), ctrW, col[5][1])
               for col in _RV_GRID])

    # ---- data: two physical rows per record, whole-record alternating band ----
    for n, key in enumerate(union):
        if events is not None and n and n % _PROGRESS_EVERY == 0:
            events.on_log(f"  Report View: {n:,} of {len(union):,} records…")
        ra = rows_a[amap[key]] if key in amap else None
        rb = rows_b[bmap[key]] if key in bmap else None
        one = (tsn_one[bmap[key]] if (tsn_one and key in bmap and bmap[key] < len(tsn_one)) else {})
        location = ""
        if key in amap and amap[key] < len(tm_loc) and tm_loc[amap[key]]:
            location = tm_loc[amap[key]]
        elif one:
            location = one.get("LOC", "")
        pmval = aval(ra, "PM") or aval(rb, "PM") or (key[1] if len(key) > 1 else "")
        alt = (n % 2 == 1)
        # A location present in only ONE system isn't a row of field conflicts — mirror the
        # Comparison sheet's "Only in TSMIS/TSN": show the present side's values as a solid
        # side-colored band (orange=TSMIS-only via 'tm', blue=TSN-only via 'tn'), label it in
        # the count columns ("TSMIS"/"TSN" + "only"), and keep it OUT of the Major/Diffs tally
        # (a presence gap, not field disagreements). Identity repeats on both physical rows so
        # a filter keeps the 2-row record intact.
        if ra is None or rb is None:
            side = "tm" if rb is None else "tn"          # tm = TSMIS-only, tn = TSN-only
            present = ra if rb is None else rb
            label = sc.side_a if rb is None else sc.side_b
            for li in (0, 1):
                bottom = (li == 1)
                row = [woc(label, side, alt, bottom=bottom),
                       woc("only", side, alt, bottom=bottom),
                       woc(key[0], side, alt, bottom=bottom)]
                for col in _RV_GRID:
                    spec = col[2] if li == 0 else col[5]
                    kind, ref = spec
                    if kind == "loc":
                        text = location
                    elif kind == "pm":
                        text = pmval
                    elif kind == "blank":
                        text = ""
                    elif kind == "tn":                   # TSN-only source column
                        v = one.get(ref, "") if (side == "tn" and one) else ""
                        text = _iso_date(v) if ref in _RV_DATEONE else v
                    else:                                # 'cmp'/'tm' -> the present side's value
                        text = aval(present, ref)
                    align = lft if (li == 1 and col[4] == "DESCRIPTION") else None
                    row.append(woc(text, side, alt, bottom=bottom, align=align))
                ws.append(row)
            continue
        # F1: evaluate each grid cell ONCE (it used to run twice — a counting
        # pass then a rendering pass), count from the cached results, render from
        # them. Same values, same statuses, half the evaluation work.
        vals = {(li, ci): value(col[2] if li == 0 else col[5], ra, rb, one)
                for li in (0, 1) for ci, col in enumerate(_RV_GRID)}
        maj = sum(1 for _t, st in vals.values() if st == "hard")
        dif = sum(1 for _t, st in vals.values() if st in ("soft", "hard"))
        for li in (0, 1):
            bottom = (li == 1)
            row = [woc(maj, "hard" if maj else "count", alt, bottom=bottom),
                   woc(dif, "count", alt, bottom=bottom),
                   woc(key[0], "count", alt, bottom=bottom)]
            for ci, col in enumerate(_RV_GRID):
                kind = (col[2] if li == 0 else col[5])[0]
                if kind == "loc":
                    text, st = location, "id"
                elif kind == "pm":
                    text, st = pmval, "id"
                else:
                    text, st = vals[(li, ci)]
                align = lft if (li == 1 and col[4] == "DESCRIPTION") else None
                row.append(woc(text, st, alt, bottom=bottom, align=align))
            ws.append(row)

    # ---- header merges (aux labels down rows 1-4; group runs across) ----
    for i in range(1, NA + 1):
        ws.merged_cells.ranges.add(CellRange(min_col=i, max_col=i, min_row=1, max_row=4))

    def merge_groups(groups, hdr_row):
        i = 0
        while i < NG:
            g, j = groups[i], i
            while j < NG and groups[j] == g:
                j += 1
            if g and j - i > 1:
                ws.merged_cells.ranges.add(CellRange(
                    min_col=NA + i + 1, max_col=NA + j, min_row=hdr_row, max_row=hdr_row))
            i = j
    merge_groups(g1, 1)
    merge_groups(g2, 3)

    ws.auto_filter.ref = f"A4:{get_column_letter(NC)}{4 + 2 * len(union)}"
    for h, ht in {1: 13, 2: 22, 3: 13, 4: 14}.items():
        ws.row_dimensions[h].height = ht
    WG = {"SFX": 5, "P": 3.5, "POST MILE": 8, "S": 3, "LOCATION": 13, "DATE OF REC": 9.5,
          "H/G": 4, "CITY": 6, "R/U": 5, "N/L": 6}
    for ci, w in {1: 5.5, 2: 5.5, 3: 8}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    for gi, col in enumerate(_RV_GRID):
        lab = col[1]
        w = (10.5 if lab in ("EFF-DATE", "ML 2nd EFF") else WG[lab] if lab in WG
             else 9 if lab == "ML ADT" else 4.2)
        ws.column_dimensions[get_column_letter(NA + gi + 1)].width = w
    return ws


# --------------------------------------------------------------------------- #
# adapter surface
# --------------------------------------------------------------------------- #
def suggest_name(tsmis_path):
    return suggest_route_name(tsmis_path, "Intersection_Detail",
                              "TSMIS_vs_TSN_IntersectionDetail")


def _load_pair(tsmis_path, tsn_path):
    """(rows_t, rows_n, warnings) for the shared driver — no input warnings on this
    FLAT detail pair, so run_compare uses its () default."""
    rows_t, _ = _load_tsmis(tsmis_path)
    rows_n, _ = _load_tsn(tsn_path)
    return rows_t, rows_n, None


def compare(tsmis_path, tsn_path, out_path, events=None, confirm_overwrite=None,
            mode="formulas", commit_guard=None):
    """Build the Intersection Detail TSMIS-vs-TSN comparison workbook(s). `tsmis_path`
    is the consolidated TSMIS Intersection Detail workbook; `tsn_path` the TSN
    statewide (raw or normalized) workbook.

    A per-call schema adds the two-line 'Report View' replica via the EXISTING
    extra_sheet_writer opt-in (the flat Comparison sheet is untouched; compare_core
    stays unmodified). The TSN-only columns come from the raw TSN file (None for a
    normalized library) and the locations from the consolidated TSMIS — both read
    lazily inside the writer, so they only open the workbooks when a sheet is actually
    built (after a successful load)."""
    schema = dataclasses.replace(
        _SCHEMA,
        extra_sheet_writer=lambda wb, ctx: _write_report_view(
            wb, ctx, _tsn_onesided(Path(tsn_path)), _tsmis_locations(Path(tsmis_path))),
        report_view_diff_check=("Report View", "B", 2))
    return ctc.run_files_compare(
        schema, tsmis_path, tsn_path, out_path,
        banner="Intersection Detail Comparison — TSMIS vs TSN", has_route=True,
        loader=_load_pair, deps_ok=_DEPS_OK,
        deps_msg="Required components are missing (openpyxl).",
        events=events, confirm_overwrite=confirm_overwrite, mode=mode,
        commit_guard=commit_guard)
