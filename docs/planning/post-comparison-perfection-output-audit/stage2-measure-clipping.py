"""Stage 2 bounded recheck: is visible text clipped in the delivered workbooks?

Resolves the central Stage 1 conflict:
  * Codex PCOA-CX-013 / PCOA-CX-014 deny ~every produced deliverable because
    Summary / Spot Check / Comparison label cells are materially clipped.
  * Claude VC-14 states "presentation contracts hold ... no clipping".

The two rounds measured DIFFERENT cells, so this script measures the exact
cells Codex names, on workbooks Claude generated (a neutral artifact for this
question -- neither reviewer's visual tooling is reused).

Method, deliberately conservative:
  * a cell is only reported when it CANNOT overflow: no wrap, no shrink-to-fit,
    and the next populated cell to the right blocks Excel's normal spill;
  * merged ranges are credited with their full combined width;
  * text is measured with the real Calibri 11 glyph metrics (the workbooks'
    font) via PIL, converted to Excel's pixel model (width chars -> px);
  * a 6-pixel tolerance is applied, matching the tolerance Codex declared.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import ImageFont

DEFAULT_COL_WIDTH = 8.43        # Excel default for Calibri 11
MDW = 7                         # max digit width in px for Calibri 11
TOLERANCE_PX = 6                # same tolerance Codex declared
SHEETS = ("Summary", "Spot Check", "Comparison")
MAX_LABEL_COL = 8               # labels/instructions live in the left columns
SCAN_ROWS = 80                  # Summary/Spot Check are short; Comparison header band

_FONT_CACHE: dict[tuple[str, int], ImageFont.FreeTypeFont] = {}


def font_for(bold: bool, size_pt: float) -> ImageFont.FreeTypeFont:
    name = "calibrib.ttf" if bold else "calibri.ttf"
    px = max(1, int(round(size_pt * 96.0 / 72.0)))
    key = (name, px)
    if key not in _FONT_CACHE:
        _FONT_CACHE[key] = ImageFont.truetype(f"C:/Windows/Fonts/{name}", px)
    return _FONT_CACHE[key]


def text_px(text: str, bold: bool, size_pt: float) -> float:
    return font_for(bold, size_pt).getlength(text)


def width_to_px(width_chars: float) -> float:
    """Excel's own column-width -> pixel conversion."""
    return round(width_chars * MDW) + 5


def merged_map(ws):
    """cell coordinate -> (anchor_coord, total_px) for every merged range."""
    out = {}
    for rng in ws.merged_cells.ranges:
        total = sum(
            width_to_px(col_width(ws, c))
            for c in range(rng.min_col, rng.max_col + 1)
        )
        anchor = f"{get_column_letter(rng.min_col)}{rng.min_row}"
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                out[f"{get_column_letter(c)}{r}"] = (anchor, total)
    return out


def col_width(ws, col_idx: int) -> float:
    dim = ws.column_dimensions.get(get_column_letter(col_idx))
    if dim is not None and dim.width is not None:
        return float(dim.width)
    return DEFAULT_COL_WIDTH


def right_neighbour_blocks(ws, row: int, col: int, max_col: int) -> bool:
    """True when the immediate right cell is populated, blocking overflow."""
    if col + 1 > max_col:
        return False
    v = ws.cell(row=row, column=col + 1).value
    return v is not None and str(v).strip() != ""


def audit_sheet(ws, wb_label: str, sheet: str) -> list[dict]:
    merged = merged_map(ws)
    hits: list[dict] = []
    max_row = min(ws.max_row, SCAN_ROWS)
    max_col = min(ws.max_column, MAX_LABEL_COL)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is None or not isinstance(val, str) or not val.strip():
                continue
            coord = cell.coordinate
            align = cell.alignment
            if align is not None and (align.wrap_text or align.shrink_to_fit):
                continue
            m = merged.get(coord)
            if m is not None:
                anchor, avail_px = m
                if anchor != coord:
                    continue          # only the anchor renders the text
            else:
                avail_px = width_to_px(col_width(ws, col))
                if not right_neighbour_blocks(ws, row, col, ws.max_column):
                    continue          # text may legally spill; not clipped
            font = cell.font
            need_px = text_px(
                val, bool(font and font.bold), float(font.size or 11) if font else 11.0
            )
            # Excel pads ~2px each side inside a cell
            if need_px + 4 > avail_px + TOLERANCE_PX:
                hits.append(
                    {
                        "workbook": wb_label,
                        "sheet": sheet,
                        "cell": coord,
                        "merged": m is not None,
                        "available_px": round(avail_px, 1),
                        "needed_px": round(need_px + 4, 1),
                        "short_by_px": round(need_px + 4 - avail_px, 1),
                        "chars": len(val),
                        "text": val if len(val) <= 90 else val[:87] + "...",
                    }
                )
    return hits


def audit_workbook(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=False, data_only=True)
    hits: list[dict] = []
    for sheet in SHEETS:
        if sheet in wb.sheetnames:
            hits.extend(audit_sheet(wb[sheet], path.name, sheet))
    wb.close()
    return hits


def main(argv: list[str]) -> int:
    targets = [Path(p) for p in argv[1:]]
    report: dict[str, object] = {"tolerance_px": TOLERANCE_PX, "workbooks": {}}
    grand = 0
    for path in targets:
        if not path.exists():
            print(f"MISSING {path}")
            continue
        hits = audit_workbook(path)
        grand += len(hits)
        report["workbooks"][path.name] = hits
        print(f"\n=== {path.name} -- {len(hits)} materially clipped cells ===")
        for h in hits[:14]:
            print(
                f"  {h['sheet']:<11} {h['cell']:<6} short {h['short_by_px']:>6.1f}px "
                f"({h['chars']:>3} chars, avail {h['available_px']:.0f}px) "
                f"{'[merged] ' if h['merged'] else ''}{h['text']!r}"
            )
        if len(hits) > 14:
            print(f"  ... and {len(hits) - 14} more")
    report["total_clipped_cells"] = grand
    out = Path(__file__).with_name("witness") / "clipping_recheck.json"
    out.write_text(json.dumps(report, indent=1), encoding="utf-8")
    print(f"\nTOTAL materially clipped cells: {grand}")
    print(f"witness -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
